"""Typed ownership and selection for historical adjusted operating metrics.

The pipeline transports adjusted metrics in a wide DataFrame because one source table
often reports several metrics together.  This module is the semantic boundary that
turns those columns into independently owned facts before any per-period selection or
TTM calculation occurs.
"""
from __future__ import annotations

import hashlib
import math
import re
from dataclasses import dataclass
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook

from .cache_semantics import ADJUSTED_METRIC_HISTORY_SELECTION_VERSION
from .longitudinal_memory.identity import build_identity


ADJUSTED_METRIC_HISTORY_CONTRACT = "contract:adjusted-metric-history-ownership@1"


class AdjustedMetricHistoryError(ValueError):
    """Raised when adjusted-metric ownership cannot be resolved deterministically."""


class AdjustedMetricId(str, Enum):
    ADJUSTED_EBIT = "adj_ebit"
    ADJUSTED_EBITDA = "adj_ebitda"
    ADJUSTED_FCF = "adj_fcf"


class AdjustedMetricPeriodType(str, Enum):
    QUARTER = "quarter"
    YTD = "ytd"
    FY = "fy"
    TTM = "ttm"


class AdjustedMetricScope(str, Enum):
    REPORTED_CONSOLIDATED = "reported_consolidated_at_period"
    CONTINUING_OPERATIONS_RECAST = "continuing_operations_current_presentation"


class AdjustedMetricSourceRole(str, Enum):
    DIRECT = "direct"
    DERIVED_EXACT = "derived_exact"
    LEGACY = "legacy"


_METRIC_IDS: Tuple[AdjustedMetricId, ...] = (
    AdjustedMetricId.ADJUSTED_EBIT,
    AdjustedMetricId.ADJUSTED_EBITDA,
    AdjustedMetricId.ADJUSTED_FCF,
)

_METRIC_DISPLAY_LABEL = {
    AdjustedMetricId.ADJUSTED_EBIT: "Adjusted EBIT",
    AdjustedMetricId.ADJUSTED_EBITDA: "Adjusted EBITDA",
    AdjustedMetricId.ADJUSTED_FCF: "Adjusted FCF",
}

_REPORTED_DEFINITION_IDS = {
    AdjustedMetricId.ADJUSTED_EBIT: "definition:issuer-reported-consolidated-adjusted-ebit@1",
    AdjustedMetricId.ADJUSTED_EBITDA: "definition:issuer-reported-consolidated-adjusted-ebitda@1",
    AdjustedMetricId.ADJUSTED_FCF: "definition:issuer-reported-consolidated-adjusted-fcf@1",
}


def reported_adjusted_metric_definition_id(metric_id: AdjustedMetricId | str) -> str:
    metric = metric_id if isinstance(metric_id, AdjustedMetricId) else AdjustedMetricId(str(metric_id))
    return _REPORTED_DEFINITION_IDS[metric]


@dataclass(frozen=True)
class AdjustedMetricHistorySelection:
    selected_facts: pd.DataFrame
    quarter_values: Mapping[str, Mapping[pd.Timestamp, Optional[float]]]
    ttm_values: Mapping[str, Mapping[pd.Timestamp, Optional[float]]]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return re.sub(r"\s+", " ", str(value)).strip()


def _coerce_period_type(value: Any, *, source_column_label: str = "") -> AdjustedMetricPeriodType:
    token = _clean_text(value).lower()
    aliases = {
        "": AdjustedMetricPeriodType.QUARTER,
        "quarter": AdjustedMetricPeriodType.QUARTER,
        "3m": AdjustedMetricPeriodType.QUARTER,
        "ytd": AdjustedMetricPeriodType.YTD,
        "fy": AdjustedMetricPeriodType.FY,
        "annual": AdjustedMetricPeriodType.FY,
        "year": AdjustedMetricPeriodType.FY,
        "ttm": AdjustedMetricPeriodType.TTM,
    }
    if token in aliases:
        return aliases[token]
    column_low = source_column_label.lower()
    if "3m" in column_low or "three month" in column_low:
        return AdjustedMetricPeriodType.QUARTER
    if "ytd" in column_low or "nine month" in column_low or "six month" in column_low:
        return AdjustedMetricPeriodType.YTD
    if "twelve month" in column_low or "year ended" in column_low:
        return AdjustedMetricPeriodType.FY
    raise AdjustedMetricHistoryError(f"Unsupported adjusted-metric period type {value!r}.")


def _default_authority(row: Mapping[str, Any]) -> Tuple[str, int, AdjustedMetricSourceRole]:
    source = _clean_text(row.get("source")).lower()
    confidence = _clean_text(row.get("confidence")).lower()
    if source == "issuer_recast_workbook":
        return "issuer_recast_current_presentation", 400, AdjustedMetricSourceRole.DIRECT
    if source == "ex99":
        return (
            "issuer_direct_period_release",
            300 if confidence != "low" else 250,
            AdjustedMetricSourceRole.DIRECT,
        )
    if source:
        return (
            "registered_local_material",
            200 if confidence != "low" else 100,
            AdjustedMetricSourceRole.LEGACY,
        )
    return (
        "legacy_adjusted_metric_transport",
        75 if confidence == "high" else 50,
        AdjustedMetricSourceRole.LEGACY,
    )


def _metric_metadata(row: Mapping[str, Any], metric_id: AdjustedMetricId, suffix: str) -> Any:
    return row.get(f"{metric_id.value}_{suffix}")


def adjusted_metric_facts_from_wide_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Normalize every populated wide metric cell into one typed fact row."""

    columns = [
        "quarter",
        "metric_id",
        "value",
        "period_type",
        "basis",
        "scope",
        "definition_id",
        "source_role",
        "source_authority",
        "authority_rank",
        "source",
        "source_type",
        "source_document_id",
        "source_occurrence_id",
        "source_locator",
        "source_metric_label",
        "source_unit",
        "raw_source_scalar",
        "canonical_usd_millions",
        "doc",
        "accn",
        "page",
        "method",
        "col",
        "source_snippet",
        "note",
        "doc_type",
        "filed",
        "confidence",
        "derivation_rule",
        "derivation_input_occurrence_ids",
        "transport_row_index",
    ]
    if not isinstance(frame, pd.DataFrame) or frame.empty or "quarter" not in frame.columns:
        return pd.DataFrame(columns=columns)

    facts: List[Dict[str, Any]] = []
    for row_index, row_series in frame.iterrows():
        row = row_series.to_dict()
        quarter = pd.to_datetime(row.get("quarter"), errors="coerce")
        if pd.isna(quarter):
            continue
        quarter = pd.Timestamp(quarter).normalize()
        default_authority, default_rank, default_role = _default_authority(row)
        for metric_id in _METRIC_IDS:
            numeric = pd.to_numeric(pd.Series([row.get(metric_id.value)]), errors="coerce").iloc[0]
            if pd.isna(numeric):
                continue
            source_column_label = _clean_text(
                _metric_metadata(row, metric_id, "source_column_label") or row.get("col")
            )
            period_type = _coerce_period_type(
                _metric_metadata(row, metric_id, "period_type") or row.get("period_type"),
                source_column_label=source_column_label,
            )
            scope_raw = _clean_text(_metric_metadata(row, metric_id, "scope"))
            scope = scope_raw or AdjustedMetricScope.REPORTED_CONSOLIDATED.value
            definition_id = _clean_text(_metric_metadata(row, metric_id, "definition_id"))
            if not definition_id:
                definition_id = _REPORTED_DEFINITION_IDS[metric_id]
            basis = _clean_text(_metric_metadata(row, metric_id, "basis")) or "adjusted_non_gaap"
            authority = _clean_text(_metric_metadata(row, metric_id, "source_authority")) or default_authority
            authority_rank_raw = _metric_metadata(row, metric_id, "authority_rank")
            authority_rank = int(authority_rank_raw) if pd.notna(authority_rank_raw) else int(default_rank)
            role = _clean_text(_metric_metadata(row, metric_id, "source_role")) or default_role.value
            source_document_id = _clean_text(
                _metric_metadata(row, metric_id, "source_document_id") or row.get("source_document_id")
            )
            facts.append(
                {
                    "quarter": quarter,
                    "metric_id": metric_id.value,
                    "value": float(numeric),
                    "period_type": period_type.value,
                    "basis": basis,
                    "scope": scope,
                    "definition_id": definition_id,
                    "source_role": role,
                    "source_authority": authority,
                    "authority_rank": authority_rank,
                    "source": _clean_text(row.get("source")),
                    "source_type": _clean_text(row.get("source_type")),
                    "source_document_id": source_document_id,
                    "source_occurrence_id": _clean_text(
                        _metric_metadata(row, metric_id, "source_occurrence_id")
                    ),
                    "source_locator": _clean_text(_metric_metadata(row, metric_id, "source_locator")),
                    "source_metric_label": _clean_text(
                        _metric_metadata(row, metric_id, "source_metric_label")
                    )
                    or _METRIC_DISPLAY_LABEL[metric_id],
                    "source_unit": _clean_text(_metric_metadata(row, metric_id, "source_scale")),
                    "raw_source_scalar": _metric_metadata(row, metric_id, "raw_source_scalar"),
                    "canonical_usd_millions": _metric_metadata(
                        row, metric_id, "canonical_usd_millions"
                    ),
                    "doc": _clean_text(row.get("doc")),
                    "accn": _clean_text(row.get("accn")),
                    "page": row.get("page"),
                    "method": _clean_text(row.get("method")),
                    "col": _clean_text(row.get("col")),
                    "source_snippet": _clean_text(row.get("source_snippet")),
                    "note": _clean_text(row.get("note")),
                    "doc_type": _clean_text(row.get("doc_type")),
                    "filed": row.get("filed"),
                    "confidence": _clean_text(row.get("confidence")),
                    "derivation_rule": _clean_text(
                        _metric_metadata(row, metric_id, "derivation_rule")
                    ),
                    "derivation_input_occurrence_ids": tuple(
                        _metric_metadata(row, metric_id, "derivation_input_occurrence_ids") or ()
                    ),
                    "transport_row_index": str(row_index),
                }
            )
    return pd.DataFrame(facts, columns=columns)


def _quarter_number(period: pd.Timestamp) -> int:
    month = int(pd.Timestamp(period).month)
    if month not in {3, 6, 9, 12}:
        raise AdjustedMetricHistoryError(
            f"Adjusted-metric quarter must end in March, June, September, or December; "
            f"received {pd.Timestamp(period).date()}."
        )
    return month // 3


def _derive_quarter_facts_from_ytd(facts: pd.DataFrame) -> pd.DataFrame:
    """Create exact quarter facts from compatible YTD facts.

    A direct quarter remains the preferred owner.  Derived rows are merely eligible
    lower-ranked facts and retain both input occurrences.  The definition, scope,
    basis, metric, and calendar year must match exactly before subtraction.
    """

    if facts.empty:
        return facts
    ytd = facts[facts["period_type"] == AdjustedMetricPeriodType.YTD.value].copy()
    if ytd.empty:
        return facts
    derived: List[Dict[str, Any]] = []
    group_columns = ["metric_id", "basis", "scope", "definition_id"]
    for semantic_key, group in ytd.groupby(group_columns, sort=True, dropna=False):
        by_year: Dict[int, Dict[int, Dict[str, Any]]] = {}
        for record in group.to_dict("records"):
            period = pd.Timestamp(record["quarter"]).normalize()
            quarter_number = _quarter_number(period)
            year_records = by_year.setdefault(int(period.year), {})
            existing = year_records.get(quarter_number)
            if existing is not None:
                same_value = math.isclose(
                    float(existing["value"]),
                    float(record["value"]),
                    rel_tol=0.0,
                    abs_tol=1e-6,
                )
                if not same_value:
                    raise AdjustedMetricHistoryError(
                        f"Conflicting YTD adjusted-metric facts for {semantic_key!r} at "
                        f"{period.date()}: {[existing['value'], record['value']]!r}."
                    )
                if int(record["authority_rank"]) > int(existing["authority_rank"]):
                    year_records[quarter_number] = record
                elif int(record["authority_rank"]) == int(existing["authority_rank"]):
                    year_records[quarter_number] = min(
                        (existing, record), key=_stable_fact_key
                    )
            else:
                year_records[quarter_number] = record

        for year, year_records in sorted(by_year.items()):
            for quarter_number, current in sorted(year_records.items()):
                if quarter_number == 1:
                    inputs = [current]
                    quarter_value = float(current["value"])
                    rule = "derivation:adjusted-metric-q1-ytd-equals-quarter@1"
                else:
                    prior = year_records.get(quarter_number - 1)
                    if prior is None:
                        continue
                    inputs = [prior, current]
                    quarter_value = float(current["value"]) - float(prior["value"])
                    rule = "derivation:adjusted-metric-ytd-minus-prior-ytd@1"
                input_ids = tuple(
                    _clean_text(record.get("source_occurrence_id"))
                    for record in inputs
                )
                if any(not item for item in input_ids):
                    continue
                period = pd.Timestamp(current["quarter"]).normalize()
                occurrence_id = build_identity(
                    "adjusted-metric-derived-occurrence",
                    (
                        ("metric", str(current["metric_id"])),
                        ("period", str(period.date())),
                        ("rule", rule),
                        ("inputs", "|".join(input_ids)),
                    ),
                )
                derived_record = dict(current)
                derived_record.update(
                    {
                        "value": quarter_value,
                        "period_type": AdjustedMetricPeriodType.QUARTER.value,
                        "source_role": AdjustedMetricSourceRole.DERIVED_EXACT.value,
                        "source_authority": "issuer_exact_ytd_residual",
                        "authority_rank": max(0, min(int(item["authority_rank"]) for item in inputs) - 1),
                        "source_occurrence_id": occurrence_id,
                        "source_locator": rule,
                        "raw_source_scalar": None,
                        "canonical_usd_millions": quarter_value / 1_000_000.0,
                        "derivation_rule": rule,
                        "derivation_input_occurrence_ids": input_ids,
                        "transport_row_index": "",
                    }
                )
                derived.append(derived_record)
    if not derived:
        return facts
    return pd.concat([facts, pd.DataFrame(derived)], ignore_index=True, sort=False)


def _semantic_key(row: Mapping[str, Any]) -> Tuple[str, str, str, str]:
    return (
        _clean_text(row.get("period_type")),
        _clean_text(row.get("basis")),
        _clean_text(row.get("scope")),
        _clean_text(row.get("definition_id")),
    )


def _stable_fact_key(row: Mapping[str, Any]) -> Tuple[str, ...]:
    return (
        _clean_text(row.get("source_document_id")),
        _clean_text(row.get("source_occurrence_id")),
        _clean_text(row.get("source_locator")),
        _clean_text(row.get("doc")),
        _clean_text(row.get("source_authority")),
    )


def select_adjusted_metric_history(frame: pd.DataFrame) -> pd.DataFrame:
    """Select one source-order-independent owner for each period and metric."""

    facts = _derive_quarter_facts_from_ytd(adjusted_metric_facts_from_wide_frame(frame))
    if facts.empty:
        return facts
    selected: List[Dict[str, Any]] = []
    group_cols = ["quarter", "metric_id"]
    for (quarter, metric_id), group in facts.groupby(group_cols, sort=True, dropna=False):
        quarter_group = group[group["period_type"] == AdjustedMetricPeriodType.QUARTER.value].copy()
        if quarter_group.empty:
            continue
        top_rank = int(quarter_group["authority_rank"].max())
        top = quarter_group[quarter_group["authority_rank"] == top_rank].copy()
        semantic_keys = {_semantic_key(row) for row in top.to_dict("records")}
        if len(semantic_keys) != 1:
            raise AdjustedMetricHistoryError(
                f"Conflicting adjusted-metric definitions for {metric_id} at "
                f"{pd.Timestamp(quarter).date()}: {sorted(semantic_keys)!r}."
            )
        values = [float(v) for v in top["value"].tolist()]
        first_value = values[0]
        if any(not math.isclose(value, first_value, rel_tol=0.0, abs_tol=1e-6) for value in values[1:]):
            owners = sorted(_stable_fact_key(row) for row in top.to_dict("records"))
            raise AdjustedMetricHistoryError(
                f"Conflicting adjusted-metric values for {metric_id} at "
                f"{pd.Timestamp(quarter).date()}: {values!r}; owners={owners!r}."
            )
        records = top.to_dict("records")
        records.sort(key=_stable_fact_key)
        chosen = dict(records[0])
        chosen.pop("transport_row_index", None)
        chosen["corroboration_count"] = len(records)
        selected.append(chosen)
    out = pd.DataFrame(selected)
    if out.empty:
        return out
    return out.sort_values(["quarter", "metric_id"], kind="stable").reset_index(drop=True)


def build_adjusted_metric_history_selection(
    frame: pd.DataFrame,
    quarter_list: Sequence[pd.Timestamp],
) -> AdjustedMetricHistorySelection:
    """Build quarter and compatible four-quarter TTM maps from typed facts."""

    selected = select_adjusted_metric_history(frame)
    quarters = [pd.Timestamp(q).normalize() for q in quarter_list]
    quarter_values: Dict[str, Dict[pd.Timestamp, Optional[float]]] = {}
    ttm_values: Dict[str, Dict[pd.Timestamp, Optional[float]]] = {}
    for metric_id in _METRIC_IDS:
        metric_rows = selected[selected.get("metric_id") == metric_id.value].copy() if not selected.empty else pd.DataFrame()
        by_quarter = {
            pd.Timestamp(row["quarter"]).normalize(): row
            for row in metric_rows.to_dict("records")
        }
        quarter_values[metric_id.value] = {
            q: (float(by_quarter[q]["value"]) if q in by_quarter else None)
            for q in quarters
        }
        metric_ttm: Dict[pd.Timestamp, Optional[float]] = {}
        for idx, q in enumerate(quarters):
            if idx < 3:
                metric_ttm[q] = None
                continue
            window = quarters[idx - 3 : idx + 1]
            period_ordinals = [pd.Period(window_q, freq="Q").ordinal for window_q in window]
            if any(
                current != previous + 1
                for previous, current in zip(period_ordinals, period_ordinals[1:])
            ):
                metric_ttm[q] = None
                continue
            rows = [by_quarter.get(window_q) for window_q in window]
            if any(row is None for row in rows):
                metric_ttm[q] = None
                continue
            semantic_keys = {_semantic_key(row or {}) for row in rows}
            if len(semantic_keys) != 1:
                metric_ttm[q] = None
                continue
            if any(_clean_text((row or {}).get("period_type")) != AdjustedMetricPeriodType.QUARTER.value for row in rows):
                metric_ttm[q] = None
                continue
            metric_ttm[q] = float(sum(float((row or {})["value"]) for row in rows))
        ttm_values[metric_id.value] = metric_ttm
    return AdjustedMetricHistorySelection(
        selected_facts=selected,
        quarter_values=quarter_values,
        ttm_values=ttm_values,
    )


_RECAST_FILE_PERIOD_RE = re.compile(r"\bQ([1-4])\s+(20\d{2})\b", re.IGNORECASE)


def resolve_latest_issuer_recast_workbook(directory: Path) -> Optional[Path]:
    """Resolve the latest explicitly period-labelled issuer recast workbook."""

    directory = Path(directory)
    if not directory.is_dir():
        return None
    candidates: List[Tuple[int, int, Path, str]] = []
    for path in sorted(directory.glob("*.xlsx"), key=lambda item: item.name.lower()):
        match = _RECAST_FILE_PERIOD_RE.search(path.name)
        if not match or not path.is_file():
            continue
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        candidates.append((int(match.group(2)), int(match.group(1)), path, digest))
    if not candidates:
        return None
    latest_period = max((year, quarter) for year, quarter, _path, _digest in candidates)
    latest = [item for item in candidates if item[:2] == latest_period]
    digests = {digest for _year, _quarter, _path, digest in latest}
    if len(digests) != 1:
        raise AdjustedMetricHistoryError(
            f"Ambiguous issuer recast workbooks for {latest_period[0]}-Q{latest_period[1]}: "
            f"{[str(path) for _year, _quarter, path, _digest in latest]!r}."
        )
    return sorted((path for _year, _quarter, path, _digest in latest), key=lambda item: item.name.lower())[0]


def _quarter_from_recast_header(value: Any) -> Optional[pd.Timestamp]:
    text = _clean_text(value)
    match = re.fullmatch(r"(Mar|Jun|Sep|Dec)\s+(20\d{2})", text, re.IGNORECASE)
    if not match:
        return None
    month = {"mar": 3, "jun": 6, "sep": 9, "dec": 12}[match.group(1).lower()]
    return (pd.Timestamp(year=int(match.group(2)), month=month, day=1) + pd.offsets.MonthEnd(0)).normalize()


def load_issuer_recast_adjusted_metric_history(path: Path) -> pd.DataFrame:
    """Read direct company adjusted EBIT/EBITDA from an issuer recast workbook."""

    path = Path(path)
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise AdjustedMetricHistoryError(f"Unable to open issuer recast workbook {path}: {exc}") from exc
    try:
        if "Note" not in workbook.sheetnames or "Non-GAAP rec" not in workbook.sheetnames:
            raise AdjustedMetricHistoryError(
                f"Issuer recast workbook {path} lacks Note or Non-GAAP rec ownership surfaces."
            )
        note_sheet = workbook["Note"]
        note_text = " ".join(
            _clean_text(cell.value)
            for row in note_sheet.iter_rows()
            for cell in row
            if _clean_text(cell.value)
        ).lower()
        if "continuing operations basis" not in note_text or "recast to conform to the current period presentation" not in note_text:
            raise AdjustedMetricHistoryError(
                f"Issuer workbook {path} does not explicitly own a continuing-operations current-presentation recast."
            )

        sheet = workbook["Non-GAAP rec"]
        quarter_columns: Dict[int, pd.Timestamp] = {}
        for row_index in range(1, min(int(sheet.max_row or 0), 12) + 1):
            candidates = {
                column_index: quarter
                for column_index in range(1, int(sheet.max_column or 0) + 1)
                if (quarter := _quarter_from_recast_header(sheet.cell(row_index, column_index).value)) is not None
            }
            if len(candidates) >= 2:
                quarter_columns = candidates
                break
        if not quarter_columns:
            raise AdjustedMetricHistoryError(f"Issuer recast workbook {path} has no quarter-owned columns.")

        unit_declarations = [
            _clean_text(sheet.cell(row_index, column_index).value)
            for row_index in range(1, min(int(sheet.max_row or 0), 10) + 1)
            for column_index in range(1, min(int(sheet.max_column or 0), 6) + 1)
            if "million" in _clean_text(sheet.cell(row_index, column_index).value).lower()
        ]
        if len(set(unit_declarations)) != 1:
            raise AdjustedMetricHistoryError(
                f"Issuer recast workbook {path} lacks one unambiguous USD-millions declaration: {unit_declarations!r}."
            )
        unit_declaration = unit_declarations[0]

        metric_rows: Dict[AdjustedMetricId, int] = {}
        for metric_id, expected_label in (
            (AdjustedMetricId.ADJUSTED_EBIT, "adjusted ebit"),
            (AdjustedMetricId.ADJUSTED_EBITDA, "adjusted ebitda"),
        ):
            matches = []
            for row_index in range(1, int(sheet.max_row or 0) + 1):
                label = _clean_text(sheet.cell(row_index, 2).value).lower()
                if label == expected_label:
                    matches.append(row_index)
            if len(matches) != 1:
                raise AdjustedMetricHistoryError(
                    f"Issuer recast workbook {path} has {len(matches)} exact {expected_label!r} rows."
                )
            metric_rows[metric_id] = matches[0]

        period_match = _RECAST_FILE_PERIOD_RE.search(path.name)
        definition_version = (
            f"{int(period_match.group(2))}-Q{int(period_match.group(1))}"
            if period_match
            else digest[:12]
        )
        source_document_id = build_identity(
            "issuer-workbook",
            (("name", path.name), ("sha256", digest)),
        )
        rows: List[Dict[str, Any]] = []
        for column_index, quarter in sorted(quarter_columns.items(), key=lambda item: item[1]):
            row: Dict[str, Any] = {
                "quarter": quarter,
                "period_type": AdjustedMetricPeriodType.QUARTER.value,
                "source": "issuer_recast_workbook",
                "source_type": "historical_recast",
                "doc": str(path),
                "confidence": "high",
                "source_document_id": source_document_id,
                "source_lineage_contract": ADJUSTED_METRIC_HISTORY_CONTRACT,
            }
            populated = False
            for metric_id, row_index in metric_rows.items():
                raw_value = sheet.cell(row_index, column_index).value
                numeric = pd.to_numeric(pd.Series([raw_value]), errors="coerce").iloc[0]
                if pd.isna(numeric):
                    continue
                populated = True
                coordinate = sheet.cell(row_index, column_index).coordinate
                definition_id = (
                    f"definition:issuer-continuing-operations-current-presentation-"
                    f"{metric_id.value}@{definition_version}"
                )
                occurrence_id = build_identity(
                    "adjusted-metric-occurrence",
                    (
                        ("doc", source_document_id),
                        ("sheet", "Non-GAAP rec"),
                        ("cell", coordinate),
                        ("metric", metric_id.value),
                        ("period", str(quarter.date())),
                    ),
                )
                source_millions = Decimal(str(numeric)).quantize(Decimal("0.000001"))
                canonical_value = float(source_millions * Decimal("1000000"))
                row[metric_id.value] = canonical_value
                row.update(
                    {
                        f"{metric_id.value}_metric_id": metric_id.value,
                        f"{metric_id.value}_period_type": AdjustedMetricPeriodType.QUARTER.value,
                        f"{metric_id.value}_basis": "adjusted_non_gaap",
                        f"{metric_id.value}_scope": AdjustedMetricScope.CONTINUING_OPERATIONS_RECAST.value,
                        f"{metric_id.value}_definition_id": definition_id,
                        f"{metric_id.value}_source_role": AdjustedMetricSourceRole.DIRECT.value,
                        f"{metric_id.value}_source_authority": "issuer_recast_current_presentation",
                        f"{metric_id.value}_authority_rank": 400,
                        f"{metric_id.value}_source_document_id": source_document_id,
                        f"{metric_id.value}_source_occurrence_id": occurrence_id,
                        f"{metric_id.value}_source_locator": f"sheet:Non-GAAP rec;cell:{coordinate}",
                        f"{metric_id.value}_source_metric_label": _METRIC_DISPLAY_LABEL[metric_id],
                        f"{metric_id.value}_raw_source_scalar": float(source_millions),
                        f"{metric_id.value}_source_currency": "USD",
                        f"{metric_id.value}_source_scale": "millions",
                        f"{metric_id.value}_source_scale_factor": 1_000_000.0,
                        f"{metric_id.value}_source_unit_declaration": unit_declaration,
                        f"{metric_id.value}_canonical_currency": "USD",
                        f"{metric_id.value}_canonical_value": canonical_value,
                        f"{metric_id.value}_canonical_usd_millions": float(source_millions),
                    }
                )
            if populated:
                rows.append(row)
        return pd.DataFrame(rows)
    finally:
        workbook.close()


def load_registered_issuer_recast_adjusted_metric_history(directory: Path) -> pd.DataFrame:
    path = resolve_latest_issuer_recast_workbook(directory)
    if path is None:
        return pd.DataFrame()
    return load_issuer_recast_adjusted_metric_history(path)
