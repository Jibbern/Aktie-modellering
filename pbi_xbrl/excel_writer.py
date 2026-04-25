"""Workbook save, audit-sheet writeback, and saved-workbook validation helpers."""
from __future__ import annotations

import hashlib
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .excel_writer_context import build_writer_context
from .excel_writer_core import finalize_workbook, prepare_writer_inputs, timed_writer_stage, write_qa_sheets, write_raw_data_sheets
from .excel_writer_drivers import write_driver_sheets
from .excel_writer_financials import write_debt_sheets, write_report_sheets, write_summary_sheets, write_valuation_sheets
from .excel_writer_ui import write_ui_debug_sheets, write_ui_sheets
from .market_data.service import persist_gpre_current_qtd_snapshot_history
from .pipeline_types import WorkbookInputs


@dataclass
class WorkbookWriteResult:
    saved_temp_path: Path
    quarter_notes_ui_snapshot: Dict[str, List[Tuple[str, str]]]
    quarter_notes_audit_rows: List[Dict[str, Any]] = field(default_factory=list)
    quarter_notes_header_text: str = ""
    summary_export_expectation: Dict[str, Any] = field(default_factory=dict)
    valuation_export_expectation: Dict[str, Any] = field(default_factory=dict)
    qa_export_expectation: Dict[str, Any] = field(default_factory=dict)
    needs_review_export_expectation: Dict[str, Any] = field(default_factory=dict)
    saved_workbook_provenance: Dict[str, Any] = field(default_factory=dict)


def _normalize_qnote_cell(value: Any) -> str:
    text = re.sub(r"^\[[A-Z]+\]\s*", "", str(value or "").strip())
    return re.sub(r"\s+", " ", text)


def _sanitize_excel_sheet_text(value: Any) -> Any:
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    text = value.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)
    return text


def _canonicalize_qnote_audit_text(value: Any) -> str:
    text = _normalize_qnote_cell(value)
    if not text:
        return ""

    def _collapse_leading_ngram(part_in: str) -> str:
        words = part_in.split()
        if len(words) < 4:
            return part_in
        lowered = [w.lower() for w in words]
        for n in range(min(6, len(words) // 2), 1, -1):
            prefix = lowered[:n]
            idx = n
            repeats = 1
            while idx + n <= len(words) and lowered[idx : idx + n] == prefix:
                repeats += 1
                idx += n
            if repeats > 1:
                return " ".join(words[:n] + words[idx:])
        return part_in

    parts = []
    for raw_part in re.split(r"\s*\|\s*", text):
        part = _collapse_leading_ngram(_normalize_qnote_cell(raw_part))
        if part:
            parts.append(part)
    deduped_parts: List[str] = []
    for idx, part in enumerate(parts):
        low = part.lower()
        if any(idx != jdx and other.lower() in low and len(part) >= len(other) + 12 for jdx, other in enumerate(parts)):
            continue
        if low not in {p.lower() for p in deduped_parts}:
            deduped_parts.append(part)
    text = " | ".join(deduped_parts) if deduped_parts else text
    text = re.sub(r"\b([A-Za-z][A-Za-z0-9/%$().,\- ]{3,90}?)\s+\1\b", r"\1", text, flags=re.I)
    text = _collapse_leading_ngram(text)
    return _normalize_qnote_cell(text)


def _quarter_notes_audit_cleanup_key(row: Dict[str, Any]) -> Tuple[str, str, str]:
    quarter = str(row.get("quarter") or "").strip()
    idea_label = _normalize_qnote_cell(
        row.get("idea_label") or row.get("metric_display") or row.get("family") or row.get("candidate_type") or ""
    ).lower()
    canonical_source_group = str(row.get("canonical_source_group") or "").strip().lower()
    if not canonical_source_group:
        source_type = str(row.get("source_type") or "").strip().lower()
        source_doc = str(row.get("source_doc") or "").strip()
        try:
            source_doc = Path(source_doc).name if source_doc else ""
        except Exception:
            source_doc = source_doc.replace("\\", "/").split("/")[-1]
        source_doc = re.sub(r"\.[a-z0-9]+$", "", source_doc.lower())
        source_doc = re.sub(r"^doc_\d+_", "", source_doc)
        source_doc = re.sub(r"[-_]+", "_", source_doc)
        canonical_source_group = f"{source_type or 'source'}:{source_doc or 'unknown'}"
    return (quarter, idea_label, canonical_source_group)


def _is_blob_like_qnote_audit_text(value: Any) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    if re.search(r"<[^>]+>", text):
        return True
    noisy_tokens = (
        "us-gaap:",
        "dei:",
        "xbrli:",
        "contextref",
        "xmlns",
        "link:schema",
        "ix:nonnumeric",
        "ix:nonfraction",
        "textblock",
    )
    return any(token in text for token in noisy_tokens)


def _cleanup_enriched_quarter_notes_audit_rows(
    rows_in: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    if not rows_in:
        return []
    verified_keys = {
        _quarter_notes_audit_cleanup_key(row)
        for row in rows_in
        if str(row.get("stage") or "").strip().lower() == "readback_verified"
    }
    rows_out: List[Dict[str, Any]] = []
    seen: set[Tuple[str, str, str, str, str]] = set()
    for row in rows_in:
        stage_low = str(row.get("stage") or "").strip().lower()
        cleanup_key = _quarter_notes_audit_cleanup_key(row)
        normalized_detail = _canonicalize_qnote_audit_text(row.get("final_summary") or row.get("source_excerpt") or "")
        if stage_low == "saved_workbook_missing" and cleanup_key in verified_keys:
            continue
        if (
            stage_low in {"source_detected", "candidate_created", "saved_workbook_missing"}
            and cleanup_key in verified_keys
            and _is_blob_like_qnote_audit_text(row.get("source_excerpt") or row.get("final_summary"))
        ):
            continue
        dedupe_key = (
            cleanup_key[0],
            stage_low,
            cleanup_key[1],
            cleanup_key[2],
            normalized_detail,
        )
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        rows_out.append(dict(row))
    return rows_out


def _quarter_notes_ui_snapshot_from_ws(ws: Any) -> Dict[str, List[Tuple[str, str]]]:
    rows_by_quarter: Dict[str, List[Tuple[str, str]]] = {}
    current_quarter = ""
    for rr in range(1, int(ws.max_row or 0) + 1):
        col_a = _normalize_qnote_cell(ws.cell(row=rr, column=1).value)
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", col_a):
            current_quarter = col_a
            rows_by_quarter.setdefault(current_quarter, [])
            continue
        if not current_quarter:
            continue
        category = _normalize_qnote_cell(ws.cell(row=rr, column=2).value)
        note = _normalize_qnote_cell(ws.cell(row=rr, column=3).value)
        if not note or (category.lower() == "category" and note.lower() == "note"):
            continue
        rows_by_quarter.setdefault(current_quarter, []).append((category, note))
    return rows_by_quarter


def _quarter_notes_ui_snapshot_rows_from_ws(ws: Any) -> Dict[str, List[Tuple[str, str, int]]]:
    rows_by_quarter: Dict[str, List[Tuple[str, str, int]]] = {}
    current_quarter = ""
    for rr in range(1, int(ws.max_row or 0) + 1):
        col_a = _normalize_qnote_cell(ws.cell(row=rr, column=1).value)
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", col_a):
            current_quarter = col_a
            rows_by_quarter.setdefault(current_quarter, [])
            continue
        if not current_quarter:
            continue
        category = _normalize_qnote_cell(ws.cell(row=rr, column=2).value)
        note = _normalize_qnote_cell(ws.cell(row=rr, column=3).value)
        if not note or (category.lower() == "category" and note.lower() == "note"):
            continue
        rows_by_quarter.setdefault(current_quarter, []).append((category, note, rr))
    return rows_by_quarter


def _quarter_notes_ui_header_from_ws(ws: Any) -> str:
    for coord in ("A2", "A1", "A3"):
        try:
            txt = _normalize_qnote_cell(ws[coord].value)
        except Exception:
            txt = ""
        if txt and "Generated at" in txt:
            return txt
    return _normalize_qnote_cell(ws["A1"].value)


def _cell_comment_text(cell: Any) -> str:
    comment_obj = getattr(cell, "comment", None)
    if comment_obj is None:
        return ""
    return re.sub(r"\s+", " ", str(getattr(comment_obj, "text", "") or "").strip())


def _summary_snapshot_from_ws(ws: Any) -> Dict[str, Dict[str, str]]:
    targets = [
        "What the company does",
        "Current strategic context",
        "Key competitive advantage",
    ]
    stop_labels = [
        *targets,
        "Business model / revenue streams (% of total revenue)",
        "Operating model per segment",
        "Key dependencies (3-5)",
        "What would make me wrong",
        "Key Financials",
        "Leverage / Liquidity",
        "Valuation context",
        "QA",
    ]
    out: Dict[str, Dict[str, str]] = {}
    row_lookup: Dict[str, int] = {}
    for rr in range(1, int(ws.max_row or 0) + 1):
        label = str(ws.cell(rr, 1).value or "").strip()
        if label in targets:
            row_lookup[label] = rr
    for label in targets:
        rr = row_lookup.get(label)
        if rr is None:
            continue
        parts: List[str] = []
        for nxt in range(rr + 1, int(ws.max_row or 0) + 1):
            col_a = str(ws.cell(nxt, 1).value or "").strip()
            if any(col_a == stop_label or col_a.startswith(f"{stop_label} ") or col_a.startswith(f"{stop_label} (") for stop_label in stop_labels):
                break
            col_b = str(ws.cell(nxt, 2).value or "").strip()
            if not col_a and not col_b:
                continue
            if col_a and col_b:
                parts.append(f"{col_a} {col_b}".strip())
            elif col_a:
                parts.append(col_a)
            elif col_b:
                parts.append(col_b)
        out[label] = {
            "value": "\n".join(p for p in parts if p).strip(),
            "source": _cell_comment_text(ws.cell(rr, 1)),
        }
    return out


def _valuation_snapshot_from_ws(ws: Any) -> Dict[str, Any]:
    quarter_headers: List[str] = []
    quarter_cols: List[int] = []
    for cc in range(2, int(ws.max_column or 0) + 1):
        val = str(ws.cell(6, cc).value or "").strip()
        if re.fullmatch(r"\d{4}-Q[1-4]", val):
            quarter_headers.append(val)
            quarter_cols.append(cc)
    grid_targets = [
        "Buybacks (cash)",
        "Buybacks (TTM, cash)",
        "Dividends (TTM, cash)",
        "Adj EBITDA (TTM)",
        "Adj EBIT (TTM)",
        "Net leverage",
        "Net leverage (Adj)",
        "Interest coverage (P&L TTM)",
        "Cash interest coverage (TTM)",
    ]
    hidden_targets = [
        "Buybacks (shares)",
        "Buybacks note",
        "Dividends ($/share)",
        "Dividends note",
    ]
    row_lookup: Dict[str, int] = {}
    for rr in range(1, int(ws.max_row or 0) + 1):
        label = str(ws.cell(rr, 1).value or "").strip()
        if label in set(grid_targets + hidden_targets):
            row_lookup[label] = rr
    grid_rows: Dict[str, List[Any]] = {}
    for label in grid_targets:
        rr = row_lookup.get(label)
        if rr is None:
            continue
        grid_rows[label] = [ws.cell(rr, cc).value for cc in quarter_cols]
    hidden_rows: Dict[str, Any] = {}
    for label in hidden_targets:
        rr = row_lookup.get(label)
        if rr is None:
            continue
        hidden_rows[label] = ws.cell(rr, 2).value
    return {
        "quarter_headers": quarter_headers,
        "grid_rows": grid_rows,
        "hidden_rows": hidden_rows,
    }


def _sheet_metric_rows_snapshot_from_ws(
    ws: Any,
    *,
    metrics: Optional[List[str]] = None,
    quarters: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(x or "").strip() for x in rows[0]]
    col_idx = {name: idx for idx, name in enumerate(header) if name}
    keep_cols = [col for col in ["quarter", "metric", "severity", "status", "message", "source"] if col in col_idx]
    if not keep_cols:
        return []
    metric_set = {str(x) for x in (metrics or []) if str(x).strip()}
    quarter_set = {str(x) for x in (quarters or []) if str(x).strip()}
    out: List[Dict[str, Any]] = []
    for vals in rows[1:]:
        if not vals:
            continue
        metric_val = ""
        if "metric" in col_idx and col_idx["metric"] < len(vals):
            metric_val = str(vals[col_idx["metric"]] or "").strip()
        if metric_set and metric_val not in metric_set:
            continue
        row_out: Dict[str, Any] = {}
        for col in keep_cols:
            idx = col_idx[col]
            raw_val = vals[idx] if idx < len(vals) else None
            if col == "quarter":
                qts = pd.to_datetime(raw_val, errors="coerce")
                row_out[col] = pd.Timestamp(qts).strftime("%Y-%m-%d") if pd.notna(qts) else (str(raw_val or "").strip())
            else:
                row_out[col] = "" if raw_val is None else raw_val
        if quarter_set and str(row_out.get("quarter") or "") not in quarter_set:
            continue
        if not any(str(row_out.get(col) or "").strip() for col in keep_cols if col != "quarter"):
            continue
        out.append(row_out)
    out.sort(key=lambda row: (
        str(row.get("quarter") or ""),
        str(row.get("metric") or ""),
        str(row.get("message") or ""),
    ))
    return out


def read_quarter_notes_ui_snapshot(path: Path) -> Dict[str, List[Tuple[str, str]]]:
    wb = load_workbook(Path(path), data_only=False, read_only=True)
    if "Quarter_Notes_UI" not in wb.sheetnames:
        raise RuntimeError(f"Saved workbook is missing Quarter_Notes_UI: {path}")
    try:
        return _quarter_notes_ui_snapshot_from_ws(wb["Quarter_Notes_UI"])
    finally:
        wb.close()


def read_quarter_notes_ui_header(path: Path) -> str:
    wb = load_workbook(Path(path), data_only=False, read_only=True)
    if "Quarter_Notes_UI" not in wb.sheetnames:
        raise RuntimeError(f"Saved workbook is missing Quarter_Notes_UI: {path}")
    try:
        return _quarter_notes_ui_header_from_ws(wb["Quarter_Notes_UI"])
    finally:
        wb.close()


def _file_sha1(path: Path) -> str:
    hasher = hashlib.sha1()
    with Path(path).open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _build_saved_workbook_provenance_from_wb(
    workbook_path: Path,
    wb: Any,
    *,
    qa_export_expectation: Optional[Dict[str, Any]] = None,
    needs_review_export_expectation: Optional[Dict[str, Any]] = None,
    include_sha1: bool = True,
) -> Dict[str, Any]:
    if "Quarter_Notes_UI" not in wb.sheetnames:
        raise RuntimeError(f"Saved workbook is missing Quarter_Notes_UI: {workbook_path}")
    ws = wb["Quarter_Notes_UI"]
    header_text = _quarter_notes_ui_header_from_ws(ws)
    snapshot = _quarter_notes_ui_snapshot_from_ws(ws)
    snapshot_rows = _quarter_notes_ui_snapshot_rows_from_ws(ws)
    summary_snapshot = _summary_snapshot_from_ws(wb["SUMMARY"]) if "SUMMARY" in wb.sheetnames else {}
    valuation_snapshot = _valuation_snapshot_from_ws(wb["Valuation"]) if "Valuation" in wb.sheetnames else {}
    qa_metrics = list((qa_export_expectation or {}).get("metrics") or ["QA_Buybacks"])
    qa_quarters = list((qa_export_expectation or {}).get("quarters") or [])
    needs_metrics = list((needs_review_export_expectation or {}).get("metrics") or ["buybacks_cash"])
    needs_quarters = list((needs_review_export_expectation or {}).get("quarters") or [])
    qa_checks_snapshot = (
        _sheet_metric_rows_snapshot_from_ws(wb["QA_Checks"], metrics=qa_metrics, quarters=qa_quarters)
        if "QA_Checks" in wb.sheetnames
        else []
    )
    needs_review_snapshot = (
        _sheet_metric_rows_snapshot_from_ws(wb["Needs_Review"], metrics=needs_metrics, quarters=needs_quarters)
        if "Needs_Review" in wb.sheetnames
        else []
    )
    stat = workbook_path.stat()
    return {
        "workbook_path": str(workbook_path.resolve()),
        "workbook_size": int(stat.st_size),
        "workbook_sha1": _file_sha1(workbook_path) if include_sha1 else "",
        "quarter_notes_header": header_text,
        "quarter_notes_ui_snapshot": snapshot,
        "quarter_notes_ui_snapshot_rows": snapshot_rows,
        "summary_snapshot": summary_snapshot,
        "valuation_snapshot": valuation_snapshot,
        "qa_checks_snapshot": qa_checks_snapshot,
        "needs_review_snapshot": needs_review_snapshot,
    }


def build_saved_workbook_provenance(path: Path) -> Dict[str, Any]:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        return _build_saved_workbook_provenance_from_wb(workbook_path, wb)
    finally:
        wb.close()


def validate_saved_workbook_integrity(path: Path) -> None:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    try:
        _ = wb.sheetnames
    finally:
        wb.close()
    _validate_saved_workbook_comment_xml(workbook_path)


def _validate_saved_workbook_comment_xml(workbook_path: Path) -> None:
    comment_parse_errors: List[str] = []
    with zipfile.ZipFile(workbook_path) as zf:
        for name in sorted(
            part_name
            for part_name in zf.namelist()
            if part_name.startswith("xl/comments/comment") and part_name.endswith(".xml")
        ):
            try:
                ET.fromstring(zf.read(name))
            except ET.ParseError as exc:
                comment_parse_errors.append(f"{name}: {exc}")
                if len(comment_parse_errors) >= 8:
                    break
    if comment_parse_errors:
        raise RuntimeError(
            "Saved workbook contains invalid Excel comment XML. "
            f"path={workbook_path} | " + " | ".join(comment_parse_errors)
        )


def validate_quarter_notes_ui_export(path: Path, expected_snapshot: Dict[str, List[Tuple[str, str]]]) -> None:
    actual_snapshot = read_quarter_notes_ui_snapshot(path)
    _validate_quarter_notes_ui_export_snapshot(actual_snapshot, expected_snapshot, path)


def _validate_quarter_notes_ui_export_snapshot(
    actual_snapshot: Dict[str, List[Tuple[str, str]]],
    expected_snapshot: Dict[str, List[Tuple[str, str]]],
    path: Path,
) -> None:
    expected_quarters = list(expected_snapshot.keys())
    actual_quarters = list(actual_snapshot.keys())
    if expected_quarters != actual_quarters:
        raise RuntimeError(
            "Quarter_Notes_UI export mismatch: quarter blocks differ. "
            f"expected={expected_quarters} actual={actual_quarters} path={Path(path)}"
        )
    mismatches: List[str] = []
    for quarter in expected_quarters:
        expected_rows = expected_snapshot.get(quarter, [])
        actual_rows = actual_snapshot.get(quarter, [])
        if len(expected_rows) != len(actual_rows):
            mismatches.append(
                f"{quarter}: row_count expected={len(expected_rows)} actual={len(actual_rows)}"
            )
            continue
        for idx, (expected_row, actual_row) in enumerate(zip(expected_rows, actual_rows), start=1):
            if expected_row != actual_row:
                mismatches.append(
                    f"{quarter} row {idx}: expected={expected_row!r} actual={actual_row!r}"
                )
                if len(mismatches) >= 8:
                    break
        if len(mismatches) >= 8:
            break
    if mismatches:
        raise RuntimeError(
            "Quarter_Notes_UI export mismatch after save. "
            f"path={Path(path)} | " + " | ".join(mismatches)
        )


def _norm_export_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _export_values_match(expected: Any, actual: Any, *, tolerance: float = 1e-9) -> bool:
    exp_missing = expected is None or (not isinstance(expected, str) and pd.isna(expected))
    act_missing = actual is None or actual == "" or (not isinstance(actual, str) and pd.isna(actual))
    if exp_missing and act_missing:
        return True
    if isinstance(expected, str) or isinstance(actual, str):
        return _norm_export_text(expected) == _norm_export_text(actual)
    try:
        exp_f = float(expected)
        act_f = float(actual)
    except Exception:
        return expected == actual
    return abs(exp_f - act_f) <= tolerance


def validate_summary_export(path: Path, expected_snapshot: Dict[str, Any]) -> None:
    if not expected_snapshot:
        return
    wb = load_workbook(Path(path), data_only=False, read_only=False)
    try:
        if "SUMMARY" not in wb.sheetnames:
            raise RuntimeError(f"Saved workbook is missing SUMMARY: {path}")
        actual_snapshot = _summary_snapshot_from_ws(wb["SUMMARY"])
    finally:
        wb.close()
    _validate_summary_export_snapshot(actual_snapshot, expected_snapshot, path)


def _validate_summary_export_snapshot(
    actual_snapshot: Dict[str, Any],
    expected_snapshot: Dict[str, Any],
    path: Path,
) -> None:
    mismatches: List[str] = []
    for metric, expected in dict(expected_snapshot.get("rows") or {}).items():
        actual = dict(actual_snapshot.get(metric) or {})
        if not actual:
            mismatches.append(f"{metric}: missing in saved SUMMARY")
            continue
        if not _export_values_match(expected.get("value"), actual.get("value")):
            mismatches.append(
                f"{metric}: value expected={expected.get('value')!r} actual={actual.get('value')!r}"
            )
        if not _export_values_match(expected.get("source"), actual.get("source")):
            mismatches.append(
                f"{metric}: source expected={expected.get('source')!r} actual={actual.get('source')!r}"
            )
        if len(mismatches) >= 8:
            break
    if mismatches:
        raise RuntimeError(
            "SUMMARY export mismatch after save. "
            f"path={Path(path)} | " + " | ".join(mismatches)
        )


def validate_valuation_export(path: Path, expected_snapshot: Dict[str, Any]) -> None:
    if not expected_snapshot:
        return
    wb = load_workbook(Path(path), data_only=False, read_only=True)
    try:
        if "Valuation" not in wb.sheetnames:
            raise RuntimeError(f"Saved workbook is missing Valuation: {path}")
        actual_snapshot = _valuation_snapshot_from_ws(wb["Valuation"])
    finally:
        wb.close()
    _validate_valuation_export_snapshot(actual_snapshot, expected_snapshot, path)


def _validate_valuation_export_snapshot(
    actual_snapshot: Dict[str, Any],
    expected_snapshot: Dict[str, Any],
    path: Path,
) -> None:
    mismatches: List[str] = []
    expected_headers = list(expected_snapshot.get("quarter_headers") or [])
    actual_headers = list(actual_snapshot.get("quarter_headers") or [])
    if expected_headers and expected_headers != actual_headers:
        mismatches.append(f"quarter_headers expected={expected_headers!r} actual={actual_headers!r}")
    for label, expected_vals in dict(expected_snapshot.get("grid_rows") or {}).items():
        actual_vals = list((actual_snapshot.get("grid_rows") or {}).get(label) or [])
        if not actual_vals:
            mismatches.append(f"{label}: missing in saved Valuation")
            continue
        if len(expected_vals) != len(actual_vals):
            mismatches.append(f"{label}: expected_len={len(expected_vals)} actual_len={len(actual_vals)}")
            continue
        for idx, (exp_v, act_v) in enumerate(zip(expected_vals, actual_vals), start=1):
            if not _export_values_match(exp_v, act_v):
                mismatches.append(f"{label} col {idx}: expected={exp_v!r} actual={act_v!r}")
                break
        if len(mismatches) >= 8:
            break
    for label, expected_val in dict(expected_snapshot.get("hidden_rows") or {}).items():
        actual_val = (actual_snapshot.get("hidden_rows") or {}).get(label)
        if not _export_values_match(expected_val, actual_val, tolerance=1e-6):
            mismatches.append(f"{label}: expected={expected_val!r} actual={actual_val!r}")
        if len(mismatches) >= 8:
            break
    if mismatches:
        raise RuntimeError(
            "Valuation export mismatch after save. "
            f"path={Path(path)} | " + " | ".join(mismatches)
        )


def _validate_sheet_rows_export(path: Path, expected_snapshot: Optional[Dict[str, Any]], sheet_name: str) -> None:
    if expected_snapshot is None:
        return
    expected_rows = list((expected_snapshot or {}).get("rows") or [])
    metrics = list((expected_snapshot or {}).get("metrics") or [])
    quarters = list((expected_snapshot or {}).get("quarters") or [])
    wb = load_workbook(Path(path), data_only=False, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Saved workbook is missing {sheet_name}: {path}")
        actual_rows = _sheet_metric_rows_snapshot_from_ws(wb[sheet_name], metrics=metrics, quarters=quarters)
    finally:
        wb.close()
    _validate_sheet_rows_export_snapshot(actual_rows, expected_snapshot, sheet_name, path)


def _validate_sheet_rows_export_snapshot(
    actual_rows: List[Dict[str, Any]],
    expected_snapshot: Optional[Dict[str, Any]],
    sheet_name: str,
    path: Path,
) -> None:
    mismatches: List[str] = []
    expected_rows = list((expected_snapshot or {}).get("rows") or [])
    if len(expected_rows) != len(actual_rows):
        mismatches.append(f"row_count expected={len(expected_rows)} actual={len(actual_rows)}")
    else:
        for idx, (expected_row, actual_row) in enumerate(zip(expected_rows, actual_rows), start=1):
            exp_keys = set(expected_row.keys())
            act_keys = set(actual_row.keys())
            if exp_keys != act_keys:
                mismatches.append(f"row {idx}: keys expected={sorted(exp_keys)!r} actual={sorted(act_keys)!r}")
                break
            for key in sorted(exp_keys):
                if not _export_values_match(expected_row.get(key), actual_row.get(key)):
                    mismatches.append(
                        f"row {idx} {key}: expected={expected_row.get(key)!r} actual={actual_row.get(key)!r}"
                    )
                    break
            if mismatches:
                break
    if mismatches:
        raise RuntimeError(
            f"{sheet_name} export mismatch after save. "
            f"path={Path(path)} | " + " | ".join(mismatches)
        )


def validate_qa_export(path: Path, expected_snapshot: Optional[Dict[str, Any]]) -> None:
    _validate_sheet_rows_export(path, expected_snapshot, "QA_Checks")


def validate_needs_review_export(path: Path, expected_snapshot: Optional[Dict[str, Any]]) -> None:
    _validate_sheet_rows_export(path, expected_snapshot, "Needs_Review")


def validate_saved_workbook_export(
    path: Path,
    *,
    quarter_notes_ui_snapshot: Optional[Dict[str, List[Tuple[str, str]]]] = None,
    summary_export_expectation: Optional[Dict[str, Any]] = None,
    valuation_export_expectation: Optional[Dict[str, Any]] = None,
    qa_export_expectation: Optional[Dict[str, Any]] = None,
    needs_review_export_expectation: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        if "Quarter_Notes_UI" not in wb.sheetnames:
            raise RuntimeError(f"Saved workbook is missing Quarter_Notes_UI: {workbook_path}")
        provenance = _build_saved_workbook_provenance_from_wb(
            workbook_path,
            wb,
            qa_export_expectation=qa_export_expectation,
            needs_review_export_expectation=needs_review_export_expectation,
            include_sha1=True,
        )
    finally:
        wb.close()
    _validate_saved_workbook_comment_xml(workbook_path)
    if quarter_notes_ui_snapshot is not None:
        _validate_quarter_notes_ui_export_snapshot(
            dict(provenance.get("quarter_notes_ui_snapshot") or {}),
            quarter_notes_ui_snapshot,
            workbook_path,
        )
    if summary_export_expectation:
        _validate_summary_export_snapshot(
            dict(provenance.get("summary_snapshot") or {}),
            summary_export_expectation,
            workbook_path,
        )
    if valuation_export_expectation:
        _validate_valuation_export_snapshot(
            dict(provenance.get("valuation_snapshot") or {}),
            valuation_export_expectation,
            workbook_path,
        )
    if qa_export_expectation is not None:
        _validate_sheet_rows_export_snapshot(
            list(provenance.get("qa_checks_snapshot") or []),
            qa_export_expectation,
            "QA_Checks",
            workbook_path,
        )
    if needs_review_export_expectation is not None:
        _validate_sheet_rows_export_snapshot(
            list(provenance.get("needs_review_snapshot") or []),
            needs_review_export_expectation,
            "Needs_Review",
            workbook_path,
        )
    return provenance


def validate_saved_workbook_after_audit_write(
    path: Path,
    *,
    quarter_notes_ui_snapshot: Dict[str, List[Tuple[str, str]]],
    quarter_notes_header_text: str = "",
) -> None:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    try:
        if "Quarter_Notes_Audit" not in wb.sheetnames:
            raise RuntimeError(f"Saved workbook is missing Quarter_Notes_Audit: {workbook_path}")
        if "Quarter_Notes_UI" not in wb.sheetnames:
            raise RuntimeError(f"Saved workbook is missing Quarter_Notes_UI: {workbook_path}")
        qn_ws = wb["Quarter_Notes_UI"]
        actual_header = _quarter_notes_ui_header_from_ws(qn_ws)
        actual_snapshot = _quarter_notes_ui_snapshot_from_ws(qn_ws)
    finally:
        wb.close()
    _validate_saved_workbook_comment_xml(workbook_path)
    if quarter_notes_header_text and not _export_values_match(quarter_notes_header_text, actual_header):
        raise RuntimeError(
            "Quarter_Notes_UI header mismatch after audit write. "
            f"path={workbook_path} expected={quarter_notes_header_text!r} actual={actual_header!r}"
        )
    _validate_quarter_notes_ui_export_snapshot(actual_snapshot, quarter_notes_ui_snapshot, workbook_path)


def enrich_quarter_notes_audit_rows_with_readback(
    audit_rows: List[Dict[str, Any]],
    provenance: Dict[str, Any],
) -> List[Dict[str, Any]]:
    if not audit_rows:
        return []
    snapshot_rows = provenance.get("quarter_notes_ui_snapshot_rows") or {}
    existing_trace_stages = {
        (
            str(row.get("trace_id") or "").strip(),
            str(row.get("stage") or "").strip(),
        )
        for row in audit_rows
    }
    rows_out: List[Dict[str, Any]] = []
    for row in audit_rows:
        base = dict(row)
        for key in ("workbook_path", "workbook_size", "workbook_sha1", "quarter_notes_header"):
            base[key] = provenance.get(key, base.get(key, ""))
        stage = str(base.get("stage") or "").strip()
        if stage not in {"final_selected", "routed_to_bucket"}:
            rows_out.append(base)
            continue
        quarter = str(base.get("quarter") or "")
        category = _normalize_qnote_cell(base.get("visible_category") or base.get("bucket") or "")
        note = _normalize_qnote_cell(base.get("final_summary") or "")
        if not quarter or not note:
            rows_out.append(base)
            continue
        matched_row = None
        for snap_category, snap_note, snap_row_idx in snapshot_rows.get(quarter, []):
            category_matches = (
                not category
                or stage == "routed_to_bucket"
                or _normalize_qnote_cell(snap_category) == category
            )
            if category_matches and _normalize_qnote_cell(snap_note) == note:
                matched_row = snap_row_idx
                break
        if matched_row is not None:
            trace_id = str(base.get("trace_id") or "").strip()
            if trace_id:
                for lifecycle_stage in ("source_detected", "candidate_created"):
                    if (trace_id, lifecycle_stage) in existing_trace_stages:
                        continue
                    lifecycle_row = dict(base)
                    lifecycle_row["stage"] = lifecycle_stage
                    lifecycle_row["saved_workbook_visible"] = ""
                    lifecycle_row["saved_workbook_missing"] = ""
                    lifecycle_row["saved_workbook_row"] = ""
                    lifecycle_row["readback_status"] = ""
                    lifecycle_row["attrition_class"] = ""
                    rows_out.append(lifecycle_row)
                    existing_trace_stages.add((trace_id, lifecycle_stage))
            base["stage"] = "readback_verified"
            base["saved_workbook_visible"] = True
            base["saved_workbook_missing"] = False
            base["saved_workbook_row"] = int(matched_row)
            base["readback_status"] = "verified"
            base["attrition_class"] = ""
        else:
            base["stage"] = "saved_workbook_missing"
            base["saved_workbook_visible"] = False
            base["saved_workbook_missing"] = True
            base["saved_workbook_row"] = ""
            base["readback_status"] = "missing"
            base["dropped_reason"] = str(base.get("dropped_reason") or "export_provenance_mismatch")
            if not str(base.get("attrition_class") or "").strip():
                base["attrition_class"] = "export mismatch"
        rows_out.append(base)
    return _cleanup_enriched_quarter_notes_audit_rows(rows_out)


def write_quarter_notes_audit_sheet(path: Path, audit_rows: List[Dict[str, Any]]) -> None:
    workbook_path = Path(path)
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, data_only=False, keep_vba=keep_vba)
    try:
        if "Quarter_Notes_Audit" in wb.sheetnames:
            wb.remove(wb["Quarter_Notes_Audit"])
        ws = wb.create_sheet("Quarter_Notes_Audit")
        if not audit_rows:
            ws["A1"] = "No Quarter_Notes_UI audit rows."
            wb.save(workbook_path)
            return
        preferred_cols = [
            "quarter",
            "trace_id",
            "stage",
            "source_type",
            "source_doc",
            "source_excerpt",
            "idea_label",
            "candidate_type",
            "family",
            "subject_variant",
            "bucket",
            "visible_category",
            "metric_display",
            "score_total",
            "score_components",
            "dropped_reason",
            "lost_to_trace_id",
            "merged_into_trace_id",
            "final_summary",
            "saved_workbook_visible",
            "saved_workbook_missing",
            "saved_workbook_row",
            "scope_confidence",
            "amount_confidence",
            "share_count_confidence",
            "authorization_confidence",
            "remaining_capacity_confidence",
            "dividend_change_confidence",
            "blocking_reason",
            "attrition_class",
            "workbook_path",
            "workbook_size",
            "workbook_sha1",
            "quarter_notes_header",
        ]
        extra_cols: List[str] = []
        for row in audit_rows:
            for key in row.keys():
                if key not in preferred_cols and key not in extra_cols:
                    extra_cols.append(key)
        cols = preferred_cols + extra_cols
        ws.append([_sanitize_excel_sheet_text(col) for col in cols])
        for row in audit_rows:
            row_out: List[Any] = []
            for col in cols:
                val = row.get(col, "")
                if col in {"source_excerpt", "final_summary"}:
                    val = _canonicalize_qnote_audit_text(val)
                row_out.append(_sanitize_excel_sheet_text(val))
            ws.append(row_out)
        ws.freeze_panes = "A2"
        for idx, col_name in enumerate(cols, start=1):
            width = 16.0
            if col_name in {"source_excerpt", "final_summary", "score_components", "quarter_notes_header"}:
                width = 80.0
            elif col_name in {"source_doc", "workbook_path"}:
                width = 48.0
            elif col_name in {"trace_id", "lost_to_trace_id", "merged_into_trace_id"}:
                width = 18.0
            elif col_name in {"stage", "family", "subject_variant", "candidate_type", "idea_label", "blocking_reason", "dropped_reason", "attrition_class"}:
                width = 28.0
            ws.column_dimensions[get_column_letter(idx)].width = width
        wb.save(workbook_path)
    finally:
        wb.close()


def write_excel_impl(
    out_path: Path,
    *,
    hist: pd.DataFrame,
    audit: pd.DataFrame,
    needs_review: pd.DataFrame,
    debt_tranches: pd.DataFrame,
    debt_recon: pd.DataFrame,
    adj_metrics: pd.DataFrame,
    adj_breakdown: pd.DataFrame,
    non_gaap_files: pd.DataFrame,
    adj_metrics_relaxed: pd.DataFrame,
    adj_breakdown_relaxed: pd.DataFrame,
    non_gaap_files_relaxed: pd.DataFrame,
    info_log: pd.DataFrame,
    tag_coverage: pd.DataFrame,
    period_checks: pd.DataFrame,
    qa_checks: pd.DataFrame,
    bridge_q: pd.DataFrame,
    manifest_df: pd.DataFrame,
    ocr_log: pd.DataFrame,
    qfd_preview: pd.DataFrame,
    qfd_unused: pd.DataFrame,
    debt_profile: pd.DataFrame,
    debt_tranches_latest: pd.DataFrame,
    debt_maturity: pd.DataFrame,
    debt_credit_notes: pd.DataFrame,
    revolver_df: pd.DataFrame,
    revolver_history: pd.DataFrame,
    debt_buckets: pd.DataFrame,
    slides_segments: pd.DataFrame,
    slides_debt: pd.DataFrame,
    slides_guidance: pd.DataFrame,
    quarter_notes: pd.DataFrame,
    promises: pd.DataFrame,
    promise_progress: pd.DataFrame,
    non_gaap_cred: pd.DataFrame,
    company_overview: Optional[Dict[str, Any]] = None,
    ticker: Optional[str] = None,
    price: Optional[float] = None,
    strictness: str = "ytd",
    excel_mode: str = "clean",
    is_rules: Optional[Dict[str, Any]] = None,
    cache_dir: Optional[Path] = None,
    quiet_pdf_warnings: bool = True,
    rebuild_doc_text_cache: bool = False,
    profile_timings: bool = False,
    quarter_notes_audit: bool = False,
    capture_saved_workbook_provenance: bool = True,
    excel_debug_scope: str = "full",
) -> WorkbookWriteResult:
    inputs = WorkbookInputs(
        out_path=Path(out_path),
        hist=hist,
        audit=audit,
        needs_review=needs_review,
        debt_tranches=debt_tranches,
        debt_recon=debt_recon,
        adj_metrics=adj_metrics,
        adj_breakdown=adj_breakdown,
        non_gaap_files=non_gaap_files,
        adj_metrics_relaxed=adj_metrics_relaxed,
        adj_breakdown_relaxed=adj_breakdown_relaxed,
        non_gaap_files_relaxed=non_gaap_files_relaxed,
        info_log=info_log,
        tag_coverage=tag_coverage,
        period_checks=period_checks,
        qa_checks=qa_checks,
        bridge_q=bridge_q,
        manifest_df=manifest_df,
        ocr_log=ocr_log,
        qfd_preview=qfd_preview,
        qfd_unused=qfd_unused,
        debt_profile=debt_profile,
        debt_tranches_latest=debt_tranches_latest,
        debt_maturity=debt_maturity,
        debt_credit_notes=debt_credit_notes,
        revolver_df=revolver_df,
        revolver_history=revolver_history,
        debt_buckets=debt_buckets,
        slides_segments=slides_segments,
        slides_debt=slides_debt,
        slides_guidance=slides_guidance,
        quarter_notes=quarter_notes,
        promises=promises,
        promise_progress=promise_progress,
        non_gaap_cred=non_gaap_cred,
        company_overview=company_overview,
        ticker=ticker,
        price=price,
        strictness=strictness,
        excel_mode=excel_mode,
        is_rules=is_rules,
        cache_dir=cache_dir,
        quiet_pdf_warnings=quiet_pdf_warnings,
        rebuild_doc_text_cache=rebuild_doc_text_cache,
        profile_timings=profile_timings,
        quarter_notes_audit=quarter_notes_audit,
        capture_saved_workbook_provenance=capture_saved_workbook_provenance,
        excel_debug_scope=excel_debug_scope,
    )
    return write_excel_from_inputs(inputs)


def _normalize_excel_debug_scope(scope_in: Any) -> str:
    scope_txt = str(scope_in or "full").strip().lower()
    return scope_txt if scope_txt in {"full", "drivers", "ui"} else "full"


def _ensure_partial_debug_sheet(ctx: Any, scope: str) -> None:
    if list(getattr(ctx.wb, "sheetnames", ()) or []):
        return
    ws = ctx.wb.create_sheet("Debug_Info")
    ws["A1"] = f"No visible sheets were written for excel_debug_scope={scope}."


def write_excel_from_inputs(inputs: WorkbookInputs) -> WorkbookWriteResult:
    """Render, save, and read back a workbook from one normalized input bundle.

    Expected lifecycle:
    1. Build a writer context and prepare derived/reusable writer inputs.
    2. Render visible sheets plus QA/raw-data support sheets.
    3. Save the workbook.
    4. Re-open the saved workbook and validate the delivered artifact, because the
       saved file rather than the in-memory workbook is the product truth.
    """
    excel_debug_scope = _normalize_excel_debug_scope(getattr(inputs, "excel_debug_scope", "full"))
    partial_debug_scope = excel_debug_scope != "full"
    writer_timings: Dict[str, float] = {}
    with timed_writer_stage(writer_timings, "write_excel.prep", enabled=bool(inputs.profile_timings)):
        ctx = build_writer_context(inputs)
        ctx.writer_timings = writer_timings
        ctx.state["writer_timings"] = writer_timings
        prepare_writer_inputs(ctx)

    ui_qa_rows = []
    if partial_debug_scope:
        if excel_debug_scope == "drivers":
            with timed_writer_stage(writer_timings, "write_excel.drivers", enabled=bool(inputs.profile_timings)):
                write_driver_sheets(ctx)
        elif excel_debug_scope == "ui":
            with timed_writer_stage(writer_timings, "write_excel.ui", enabled=bool(inputs.profile_timings)):
                ui_qa_rows.extend(write_ui_debug_sheets(ctx))
        _ensure_partial_debug_sheet(ctx, excel_debug_scope)
    else:
        with timed_writer_stage(writer_timings, "write_excel.summary", enabled=bool(inputs.profile_timings)):
            write_summary_sheets(ctx)
        with timed_writer_stage(writer_timings, "write_excel.valuation", enabled=bool(inputs.profile_timings)):
            ui_qa_rows.extend(write_valuation_sheets(ctx))
        with timed_writer_stage(writer_timings, "write_excel.drivers", enabled=bool(inputs.profile_timings)):
            write_driver_sheets(ctx)
        with timed_writer_stage(writer_timings, "write_excel.ui", enabled=bool(inputs.profile_timings)):
            ui_qa_rows.extend(write_ui_sheets(ctx))
        with timed_writer_stage(writer_timings, "write_excel.debt", enabled=bool(inputs.profile_timings)):
            write_debt_sheets(ctx)
        with timed_writer_stage(writer_timings, "write_excel.reports", enabled=bool(inputs.profile_timings)):
            write_report_sheets(ctx)
        with timed_writer_stage(writer_timings, "write_excel.raw_data", enabled=bool(inputs.profile_timings)):
            write_raw_data_sheets(ctx)
        with timed_writer_stage(writer_timings, "write_excel.qa", enabled=bool(inputs.profile_timings)):
            write_qa_sheets(ctx, ui_qa_rows)
    with timed_writer_stage(writer_timings, "write_excel.save", enabled=bool(inputs.profile_timings)):
        finalize_workbook(ctx)
    snapshot = (
        _quarter_notes_ui_snapshot_from_ws(ctx.wb["Quarter_Notes_UI"])
        if "Quarter_Notes_UI" in ctx.wb.sheetnames
        else {}
    )
    header_text = ""
    if "Quarter_Notes_UI" in ctx.wb.sheetnames:
        header_text = _quarter_notes_ui_header_from_ws(ctx.wb["Quarter_Notes_UI"])
    audit_rows = list(ctx.state.get("quarter_notes_audit_rows") or []) if isinstance(ctx.state, dict) else []
    if not audit_rows and isinstance(ctx.state, dict):
        ui_state = ctx.state.get("_ui_state")
        if isinstance(ui_state, dict):
            audit_rows = list(ui_state.get("quarter_notes_audit_rows") or [])
    summary_export_expectation = dict(getattr(ctx.derived, "summary_export_expectation", {}) or {})
    valuation_export_expectation = dict(getattr(ctx.derived, "valuation_export_expectation", {}) or {})
    qa_export_expectation = dict(getattr(ctx.derived, "qa_export_expectation", {}) or {})
    needs_review_export_expectation = dict(getattr(ctx.derived, "needs_review_export_expectation", {}) or {})
    saved_workbook_provenance: Dict[str, Any] = {}
    if bool(inputs.capture_saved_workbook_provenance) and not partial_debug_scope:
        try:
            saved_workbook_provenance = validate_saved_workbook_export(
                inputs.out_path,
                quarter_notes_ui_snapshot=snapshot,
                summary_export_expectation=summary_export_expectation,
                valuation_export_expectation=valuation_export_expectation,
                qa_export_expectation=qa_export_expectation,
                needs_review_export_expectation=needs_review_export_expectation,
            )
        except Exception:
            saved_workbook_provenance = {}
        if bool(inputs.quarter_notes_audit):
            try:
                audit_rows = enrich_quarter_notes_audit_rows_with_readback(audit_rows, saved_workbook_provenance)
                write_quarter_notes_audit_sheet(inputs.out_path, audit_rows)
                saved_workbook_provenance = validate_saved_workbook_export(
                    inputs.out_path,
                    quarter_notes_ui_snapshot=snapshot,
                    summary_export_expectation=summary_export_expectation,
                    valuation_export_expectation=valuation_export_expectation,
                    qa_export_expectation=qa_export_expectation,
                    needs_review_export_expectation=needs_review_export_expectation,
                )
            except Exception:
                saved_workbook_provenance = {}
    history_write_bundle = dict(ctx.state.get("gpre_current_qtd_pending_history_write") or {}) if isinstance(ctx.state, dict) else {}
    history_write_ticker_root = history_write_bundle.get("ticker_root")
    if (
        not partial_debug_scope
        and Path(inputs.out_path).exists()
        and isinstance(history_write_ticker_root, Path)
        and (
            not bool(inputs.capture_saved_workbook_provenance)
            or bool(saved_workbook_provenance)
        )
    ):
        persist_gpre_current_qtd_snapshot_history(
            history_write_ticker_root,
            history_write_bundle,
        )
    return WorkbookWriteResult(
        saved_temp_path=Path(inputs.out_path),
        quarter_notes_ui_snapshot=snapshot,
        quarter_notes_audit_rows=audit_rows,
        quarter_notes_header_text=header_text,
        summary_export_expectation=summary_export_expectation,
        valuation_export_expectation=valuation_export_expectation,
        qa_export_expectation=qa_export_expectation,
        needs_review_export_expectation=needs_review_export_expectation,
        saved_workbook_provenance=saved_workbook_provenance,
    )
