"""ANF Investment Case workbook renderers."""
from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from typing import Any, Dict, List, MutableMapping, Optional, Sequence, Set, Tuple

import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


@dataclass(frozen=True)
class AnfInvestmentCaseRenderDeps:
    runtime: MutableMapping[str, Any]


def write_anf_investment_case_data_sheet(
    deps: AnfInvestmentCaseRenderDeps,
    df: Any,
) -> None:
    runtime = deps.runtime
    wb = runtime["wb"]
    data = df
    if "ANF_Investment_Case_Data" in wb.sheetnames:
        del wb["ANF_Investment_Case_Data"]
    ws = wb.create_sheet("ANF_Investment_Case_Data")
    df = data.copy() if isinstance(data, pd.DataFrame) else pd.DataFrame()
    if df.empty:
        ws["A1"] = "No ANF investment-case data available."
        return
    preferred = [
        "section",
        "metric",
        "value",
        "unit",
        "display",
        "source",
        "source_note",
        "q1_display",
        "year_display",
        "quarter_label",
        "total_comp",
        "two_year_stack",
        "abercrombie_comp",
        "hollister_comp",
        "americas",
        "emea",
        "apac",
        "current_trend",
        "beat_miss_risk",
        "bull_evidence",
        "bear_evidence",
        "next_proof_point",
        "current_read",
        "earnings_metric",
        "multiple_yield",
        "implied_value_share",
        "scenario_read",
        "cash_flag",
        "recurring_flag",
        "quality_read",
    ]
    cols = [c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred]
    for required_col in ("section", "metric", "value", "unit", "display", "source", "source_note"):
        if required_col not in cols:
            cols.append(required_col)
        if required_col not in df.columns:
            df[required_col] = ""
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Border(left=Side(style="thin", color="D9E2EA"), right=Side(style="thin", color="D9E2EA"), top=Side(style="thin", color="D9E2EA"), bottom=Side(style="thin", color="D9E2EA"))
    for cc, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=cc, value=col)
        cell.font = Font(bold=True, color="1F2933")
        cell.fill = header_fill
        cell.border = thin
    for rr, rec in enumerate(df[cols].to_dict("records"), start=2):
        for cc, col in enumerate(cols, start=1):
            value = rec.get(col)
            if isinstance(value, str):
                value = ILLEGAL_CHARACTERS_RE.sub("", value)
            ws.cell(row=rr, column=cc, value=value)
            ws.cell(row=rr, column=cc).border = thin
            ws.cell(row=rr, column=cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=col in {"source_note", "display", "bull_evidence", "bear_evidence", "next_proof_point", "current_read", "scenario_read", "quality_read"})

    def _append_margin_bridge_audit_rows() -> None:
        if "ANF_Investment_Case" not in wb.sheetnames:
            return
        case_ws = wb["ANF_Investment_Case"]

        def _find_case_row(label: str) -> Optional[int]:
            for row_idx in range(1, int(case_ws.max_row or 0) + 1):
                if str(case_ws.cell(row_idx, 1).value or "").strip() == label:
                    return row_idx
            return None

        required_rows = {
            "revenue": _find_case_row("Forward revenue"),
            "tariff": _find_case_row("Tariff impact (bps)"),
            "freight": _find_case_row("Freight tailwind (bps)"),
            "erp": _find_case_row("ERP disruption (bps)"),
            "marketing": _find_case_row("Marketing headwind (bps)"),
            "bridge": _find_case_row("Margin bridge vs baseline"),
        }
        if any(value is None for value in required_rows.values()):
            return

        col_idx = {col: idx + 1 for idx, col in enumerate(cols)}
        value_col = col_idx.get("value")
        if not value_col:
            return

        def _case_ref(cell_ref: str) -> str:
            return f"='ANF_Investment_Case'!{cell_ref}"

        def _case_abs(col_letter: str, row_idx: Optional[int]) -> str:
            return f"${col_letter}${int(row_idx or 0)}"

        def _set(row_idx: int, col: str, value: Any) -> None:
            cc = col_idx.get(col)
            if not cc:
                return
            cell = ws.cell(row_idx, cc, value)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=col in {"source_note", "display"})
            if col == "value":
                cell.number_format = "#,##0.0"

        audit_start = int(ws.max_row or 1) + 1
        rows = [
            ("Active revenue used", _case_ref(_case_abs("G", required_rows["revenue"])), "$m", "Active Forward revenue; active cell falls back to TTM/latest-year if no manual override."),
            ("Tariff bps", _case_ref(_case_abs("G", required_rows["tariff"])), "bps", "Headwind is negative."),
            ("Freight bps", _case_ref(_case_abs("G", required_rows["freight"])), "bps", "Tailwind is positive."),
            ("ERP bps", _case_ref(_case_abs("G", required_rows["erp"])), "bps", "Headwind is negative."),
            ("Marketing bps", _case_ref(_case_abs("G", required_rows["marketing"])), "bps", "Headwind is negative."),
        ]
        for offset, (metric, formula, unit, note) in enumerate(rows):
            row_idx = audit_start + offset
            _set(row_idx, "section", "Scenario Driver Bridge Audit")
            _set(row_idx, "metric", metric)
            _set(row_idx, "value", formula)
            _set(row_idx, "unit", unit)
            _set(row_idx, "display", "")
            _set(row_idx, "source", "ANF_Investment_Case")
            _set(row_idx, "source_note", note)
        total_bps_row = audit_start + len(rows)
        value_letter = get_column_letter(value_col)
        _set(total_bps_row, "section", "Scenario Driver Bridge Audit")
        _set(total_bps_row, "metric", "Total bps")
        _set(total_bps_row, "value", f"=SUM({value_letter}{audit_start + 1}:{value_letter}{audit_start + 4})")
        _set(total_bps_row, "unit", "bps")
        _set(total_bps_row, "source", "ANF_Investment_Case")
        _set(total_bps_row, "source_note", "Tailwinds positive; headwinds negative.")

        impact_row = total_bps_row + 1
        _set(impact_row, "section", "Scenario Driver Bridge Audit")
        _set(impact_row, "metric", "Calculated $m impact")
        _set(impact_row, "value", f"={value_letter}{audit_start}*{value_letter}{total_bps_row}/10000")
        _set(impact_row, "unit", "$m")
        _set(impact_row, "source", "ANF_Investment_Case")
        _set(impact_row, "source_note", "Matches Margin bridge vs baseline active/guide formula.")

        bridge_check_row = impact_row + 1
        _set(bridge_check_row, "section", "Scenario Driver Bridge Audit")
        _set(bridge_check_row, "metric", "Bridge active / guide check")
        _set(bridge_check_row, "value", _case_ref(_case_abs("C", required_rows["bridge"])))
        _set(bridge_check_row, "unit", "$m")
        _set(bridge_check_row, "source", "ANF_Investment_Case")
        _set(bridge_check_row, "source_note", "Scenario Driver Bridge value fed into incremental EBITDA/EPS formulas.")

    _append_margin_bridge_audit_rows()
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    for cc, col in enumerate(cols, start=1):
        width = 16
        if col in {"section", "metric"}:
            width = 28
        elif col in {"display", "source_note", "bull_evidence", "bear_evidence", "next_proof_point", "current_read", "scenario_read", "quality_read"}:
            width = 44
        ws.column_dimensions[get_column_letter(cc)].width = width


def write_anf_investment_case_sheet(
    deps: AnfInvestmentCaseRenderDeps,
    df: Any,
) -> None:
    runtime = deps.runtime
    wb = runtime["wb"]
    data = df
    SCENARIO_DRIVER_CASH_FLOW_CAPEX = runtime["SCENARIO_DRIVER_CASH_FLOW_CAPEX"]
    SCENARIO_DRIVER_MARGIN_EBITDA = runtime["SCENARIO_DRIVER_MARGIN_EBITDA"]
    SCENARIO_DRIVER_SHARE_COUNT_BUYBACK = runtime["SCENARIO_DRIVER_SHARE_COUNT_BUYBACK"]
    SCENARIO_TAX_CASH_ONLY = runtime["SCENARIO_TAX_CASH_ONLY"]
    SCENARIO_TAX_NO_EPS_IMPACT = runtime["SCENARIO_TAX_NO_EPS_IMPACT"]
    SCENARIO_TAX_TAXABLE = runtime["SCENARIO_TAX_TAXABLE"]
    _ScenarioDriverBridgeSpec = runtime["_ScenarioDriverBridgeSpec"]
    _SegmentScenarioInputSpec = runtime["_SegmentScenarioInputSpec"]
    _anf_clean_visible_ui_text = runtime["_anf_clean_visible_ui_text"]
    _excel_manual_percent_active_formula = runtime["_excel_manual_percent_active_formula"]
    _excel_percent_value_expr = runtime["_excel_percent_value_expr"]
    _excel_visible_value_range_formula = runtime["_excel_visible_value_range_formula"]
    _history_q_latest_full_year_actuals_from_workbook = runtime["_history_q_latest_full_year_actuals_from_workbook"]
    _history_q_latest_full_year_period_set = runtime["_history_q_latest_full_year_period_set"]
    _history_q_year_default_formulas = runtime["_history_q_year_default_formulas"]
    _scenario_bridge_eps_value_formula = runtime["_scenario_bridge_eps_value_formula"]
    _scenario_bridge_row_values = runtime["_scenario_bridge_row_values"]
    _segment_scenario_revenue_m = runtime["_segment_scenario_revenue_m"]
    _segment_scenario_specs_from_records = runtime["_segment_scenario_specs_from_records"]
    _segment_scenario_view_basis = runtime["_segment_scenario_view_basis"]
    _write_scenario_bridge_tax_treatment_sheet = runtime["_write_scenario_bridge_tax_treatment_sheet"]
    _write_scenario_driver_assumptions_sheet = runtime["_write_scenario_driver_assumptions_sheet"]
    if "ANF_Investment_Case" in wb.sheetnames:
        del wb["ANF_Investment_Case"]
    ws = wb.create_sheet("ANF_Investment_Case")
    df = data.copy() if isinstance(data, pd.DataFrame) else pd.DataFrame()
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    ws.sheet_view.zoomScale = 112
    title_fill = PatternFill("solid", fgColor="4472C4")
    section_fill = PatternFill("solid", fgColor="5B9BD5")
    header_fill = PatternFill("solid", fgColor="EAF3F8")
    subheader_fill = PatternFill("solid", fgColor="DDEBF7")
    neutral = PatternFill("solid", fgColor="FFFFFF")
    neutral_alt = PatternFill("solid", fgColor="F7FAFC")
    callout_fill = PatternFill("solid", fgColor="F2F6FA")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Border(left=Side(style="thin", color="D9E2EA"), right=Side(style="thin", color="D9E2EA"), top=Side(style="thin", color="D9E2EA"), bottom=Side(style="thin", color="D9E2EA"))
    dark = "1F2933"
    muted = "52616B"
    visible_max_col = 10
    overflow_max_col = 11

    def _records(section: str) -> List[Dict[str, Any]]:
        if df.empty or "section" not in df.columns:
            return []
        return df[df["section"].astype(str).eq(section)].to_dict("records")

    def _rec(section: str, metric: str) -> Dict[str, Any]:
        for rec in _records(section):
            if str(rec.get("metric") or "") == metric:
                return rec
        return {}

    def _display(section: str, metric: str) -> str:
        return str(_rec(section, metric).get("display") or "")

    def _numeric_value(section: str, metric: str) -> Optional[float]:
        val = _rec(section, metric).get("value")
        try:
            if val is None or pd.isna(val):
                return None
            return float(val)
        except Exception:
            return None

    def _money_midpoint_from_display(section: str, metric: str) -> Optional[float]:
        rec = _rec(section, metric)
        if not rec:
            return None
        low = rec.get("value_low")
        high = rec.get("value_high")
        try:
            if low is not None and high is not None and not pd.isna(low) and not pd.isna(high):
                return (float(low) + float(high)) / 2.0
        except Exception:
            pass
        display = str(rec.get("display") or "")
        range_match = re.search(
            r"\$?\s*([0-9]+(?:\.[0-9]+)?)\s*(?:m|million)?\s*[-–]\s*\$?\s*([0-9]+(?:\.[0-9]+)?)\s*(?:m|million)?\b",
            display,
            flags=re.I,
        )
        if range_match:
            return (float(range_match.group(1)) + float(range_match.group(2))) / 2.0
        point_match = re.search(r"\$?\s*([0-9]+(?:\.[0-9]+)?)\s*(?:m|million)\b", display, flags=re.I)
        if point_match:
            return float(point_match.group(1))
        return None

    def _merge_row(row: int, spans: Sequence[Tuple[int, int]]) -> None:
        for first_col, last_col in spans:
            if last_col <= first_col:
                continue
            try:
                ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=last_col)
            except Exception:
                pass
            cell = ws.cell(row=row, column=first_col)
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "left", vertical="center", wrap_text=True)

    def _section(row: int, title: str, end_col: int = visible_max_col) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
        cell = ws.cell(row=row, column=1, value=title)
        cell.fill = section_fill
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for cc in range(1, end_col + 1):
            ws.cell(row=row, column=cc).fill = section_fill
            ws.cell(row=row, column=cc).border = thin
        ws.row_dimensions[row].height = 21
        return row + 1

    def _headers(
        row: int,
        labels: Sequence[str],
        end_col: Optional[int] = None,
        merge_spans: Sequence[Tuple[int, int]] = (),
    ) -> int:
        last = int(end_col or len(labels))
        for cc in range(1, last + 1):
            value = labels[cc - 1] if cc <= len(labels) else ""
            cell = ws.cell(row=row, column=cc, value=value)
            cell.fill = header_fill
            cell.font = Font(bold=True, size=12, color=dark)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center")
        _merge_row(row, merge_spans)
        ws.row_dimensions[row].height = 20
        return row + 1

    def _write_cells(
        row: int,
        vals: Sequence[Any],
        *,
        end_col: Optional[int] = None,
        fill: Optional[PatternFill] = None,
        wrap_cols: Set[int] = frozenset(),
        merge_spans: Sequence[Tuple[int, int]] = (),
    ) -> int:
        last = int(end_col or len(vals))
        row_fill = fill or neutral
        for cc in range(1, last + 1):
            value = vals[cc - 1] if cc <= len(vals) else ""
            if isinstance(value, str) and not value.startswith("="):
                value = _anf_clean_visible_ui_text(value, max_chars=220)
            cell = ws.cell(row=row, column=cc, value=value)
            cell.fill = copy(row_fill)
            cell.border = thin
            cell.font = Font(size=12, color=dark)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc in wrap_cols)
        _merge_row(row, merge_spans)
        ws.row_dimensions[row].height = 21
        return row + 1

    def _write_manual_inputs(row: int) -> Tuple[int, Dict[str, str]]:
        history_period_set = _history_q_latest_full_year_period_set(wb, ticker="ANF")
        history_fy = _history_q_latest_full_year_actuals_from_workbook(wb, ticker="ANF")
        fallback_fiscal_year = int(history_period_set.get("fiscal_year") or 2025)
        history_fy_formulas = _history_q_year_default_formulas(
            (2025, 2, 1),
            (2026, 1, 31),
            fiscal_year=fallback_fiscal_year,
            quarter_dates=history_period_set.get("quarter_dates"),
            previous_quarter_dates=history_period_set.get("previous_quarter_dates"),
            quarter_criteria=history_period_set.get("quarter_criteria"),
            previous_quarter_criteria=history_period_set.get("previous_quarter_criteria"),
            start_exclusive=True,
            end_inclusive=True,
        )

        def _fy_default(key: str) -> Any:
            val = history_fy.get(key)
            if val is None:
                return history_fy_formulas.get(key, '=""')
            if key in {"operating_margin", "tax_rate", "revenue_growth"}:
                return float(val)
            return round(float(val), 2)

        def _valuation_latest_row_value_formula(label: str) -> str:
            safe_label = str(label).replace('"', '""')
            row_expr = f'INDEX(Valuation!$B:$M,MATCH("{safe_label}",Valuation!$A:$A,0),0)'
            return f'=IFERROR(LOOKUP(2,1/({row_expr}<>""),{row_expr}),"")'

        def _numeric_or_display_midpoint(section: str, metric: str) -> Optional[float]:
            val = _numeric_value(section, metric)
            if val is not None:
                return float(val)
            display = _display(section, metric)
            nums = [
                float(match.group(0).replace(",", ""))
                for match in re.finditer(r"[-+]?\d+(?:,\d{3})*(?:\.\d+)?", str(display or ""))
            ]
            if len(nums) >= 2:
                return sum(nums[:2]) / 2.0
            if len(nums) == 1:
                return nums[0]
            return None

        def _signed_bps(value: Optional[float], sign: int) -> Any:
            if value is None:
                return '=""'
            return round(float(value) * (1 if sign >= 0 else -1), 1)

        latest_year_label = "2025 year"
        next_q_label = "2026-Q1"
        current_year_label = "2026 year"
        capex_guidance_midpoint = _money_midpoint_from_display("Guidance Beat/Miss Setup", "Capex")
        capex_guidance_display = _display("Guidance Beat/Miss Setup", "Capex")
        capex_guidance_note = (
            f"ANF capex guide midpoint from {capex_guidance_display}."
            if capex_guidance_midpoint is not None and capex_guidance_display
            else "Current-year capex guide midpoint when clean."
        )
        scenario_tax_rate = _numeric_value("Assumptions", "Tax rate")
        if scenario_tax_rate is None:
            scenario_tax_rate = history_fy.get("tax_rate")
        if scenario_tax_rate is not None and float(scenario_tax_rate) > 1.0:
            scenario_tax_rate = float(scenario_tax_rate) / 100.0
        if scenario_tax_rate is None or not (0.0 <= float(scenario_tax_rate) <= 0.35):
            scenario_tax_rate = None
        scenario_tax_note = (
            "Clean model tax rate; 25% default scenario tax rate if unavailable."
            if scenario_tax_rate is not None
            else "Uses 25% default scenario tax rate if no clean source."
        )
        buyback_actual = _numeric_value("Buybacks vs FCF", "Buybacks")
        if buyback_actual is None:
            buyback_actual = _numeric_value("Buybacks vs FCF", "2025 buybacks")
        buyback_guidance = _numeric_value("Buybacks vs FCF", "Guided buybacks")
        if buyback_guidance is None:
            buyback_guidance = 450.0
        fy_tariff_bps = _numeric_or_display_midpoint("Tariff / Margin Bridge", "2026 tariff headwind")
        q1_tariff_bps = _numeric_or_display_midpoint("Tariff / Margin Bridge", "Q1 2026 tariff headwind")
        freight_bps = _numeric_or_display_midpoint("Tariff / Margin Bridge", "Freight tailwind")
        erp_bps = _numeric_or_display_midpoint("Tariff / Margin Bridge", "ERP disruption")
        marketing_bps = _numeric_or_display_midpoint("Tariff / Margin Bridge", "Marketing")
        specs: List[Tuple[str, str, Any, Any, Any, Any, str, str]] = [
            ("price", "Current share price", '=""', '=""', '=""', '=""', "Manual input only.", "$0.00"),
            ("shares", "Diluted shares", '=IFERROR(IF(SharesDiluted<>"",SharesDiluted,Shares),"")', '=IFERROR(IF(SharesDiluted<>"",SharesDiluted,Shares),"")', '=""', '=""', "Model share denominator.", "#,##0.0"),
            ("net_debt", "Net debt / net cash", '=IFERROR(NetDebt,"")', '=IFERROR(NetDebt,"")', '=""', '=""', "Positive = net debt; negative = net cash.", "#,##0.0"),
            ("revenue", "Forward revenue", _fy_default("revenue_m"), '=IFERROR(Revenue_TTM,"")', '=""', '=""', "TTM default.", "#,##0.0"),
            ("eps", "Forward EPS", _fy_default("eps"), '=IFERROR(IF(Adj_EPS_TTM<>"",Adj_EPS_TTM,EPS_TTM),"")', "=10.6", '=""', "Adjusted EPS preferred.", "$0.00"),
            ("ebitda", "Forward Adj EBITDA", _fy_default("ebitda_m"), '=IFERROR(IF(ThesisBaseAdjEBITDA_FY<>"",ThesisBaseAdjEBITDA_FY,Adj_EBITDA),"")', '=""', '=""', "Adjusted EBITDA base.", "#,##0.0"),
            ("fcf", "Forward FCF", _fy_default("fcf_m"), '=IFERROR(IF(Adj_FCF_TTM<>"",Adj_FCF_TTM,FCF_TTM),"")', '=""', '=""', "FCF base.", "#,##0.0"),
            (
                "capex",
                "Capex",
                _fy_default("capex_m"),
                '=IFERROR(Capex_TTM,"")',
                round(float(capex_guidance_midpoint), 1) if capex_guidance_midpoint is not None else '=""',
                '=""',
                capex_guidance_note,
                "#,##0.0",
            ),
            ("pe", "P/E multiple", '=IFERROR(IF(Target_PE<>"",Target_PE,13),13)', '=IFERROR(IF(Target_PE<>"",Target_PE,13),13)', '=""', '=""', "Scenario P/E lens.", "0.0x"),
            ("ev_multiple", "EV/Adj EBITDA multiple", '=IFERROR(IF(Target_EV_AdjEBITDA<>"",Target_EV_AdjEBITDA,8),8)', '=IFERROR(IF(Target_EV_AdjEBITDA<>"",Target_EV_AdjEBITDA,8),8)', '=""', '=""', "Scenario EV/EBITDA lens.", "0.0x"),
            ("fcf_yield", "FCF yield", '=IFERROR(IF(Target_EV_Yield>1,Target_EV_Yield/100,Target_EV_Yield),0.07)', '=IFERROR(IF(Target_EV_Yield>1,Target_EV_Yield/100,Target_EV_Yield),0.07)', '=""', '=""', "Scenario FCF yield.", "0.0%"),
            ("sales_growth", "Sales growth", _fy_default("revenue_growth"), _valuation_latest_row_value_formula("Revenue YoY %"), "=0.04", '=""', "Current-year guide when clean.", "0.0%"),
            ("operating_margin", "Operating margin", _fy_default("operating_margin"), '=IFERROR(CompanyOperatingMargin_TTM,"")', "=0.1225", '=""', "12.0-12.5% guide midpoint.", "0.0%"),
            (
                "buybacks",
                "Buyback amount",
                round(float(buyback_actual), 1) if buyback_actual is not None else _valuation_latest_row_value_formula("Buybacks (TTM, cash)"),
                _valuation_latest_row_value_formula("Buybacks (TTM, cash)"),
                round(float(buyback_guidance), 1) if buyback_guidance is not None else '=""',
                '=""',
                "2026 guide; actual from capital-return readback.",
                "#,##0.0",
            ),
            ("buyback_shares", "Buyback-adjusted shares", '=IFERROR(IF(SharesDiluted<>"",SharesDiluted,Shares),"")', '=IFERROR(IF(SharesDiluted<>"",SharesDiluted,Shares),"")', '=""', '=""', "Override for post-buyback count.", "#,##0.0"),
            ("tariff_bps", "Tariff impact (bps)", '=""', '=""', _signed_bps(fy_tariff_bps, -1), _signed_bps(q1_tariff_bps, -1), "Current-year/Q1 tariff headwind.", "0"),
            ("freight_bps", "Freight tailwind (bps)", '=""', '=""', '=""', _signed_bps(freight_bps, 1), "Q1 freight tailwind.", "0"),
            ("erp_bps", "ERP disruption (bps)", '=""', '=""', '=""', _signed_bps(erp_bps, -1), "Q1 ERP headwind.", "0"),
            ("marketing_bps", "Marketing headwind (bps)", '=""', '=""', '=""', _signed_bps(marketing_bps, -1), "Q1 marketing headwind.", "0"),
            ("scenario_tax_rate", "Scenario tax rate", _fy_default("tax_rate"), scenario_tax_rate if scenario_tax_rate is not None else '=""', '=""', '=""', scenario_tax_note, "0.0%"),
        ]
        refs: Dict[str, str] = {}
        row = _section(row, "Manual Market / Scenario Inputs")
        row = _headers(
            row,
            [
                "Input",
                f"Model default ({latest_year_label})",
                "Model default (TTM)",
                f"Guidance ({current_year_label})",
                f"Guidance ({next_q_label})",
                "Manual override",
                "Active value",
                "Notes",
            ],
            end_col=visible_max_col,
            merge_spans=[(8, visible_max_col)],
        )
        for idx, (key, label, fy_default, ttm_default, current_year_guidance, next_q_guidance, note, number_format) in enumerate(specs):
            current_row = row
            active_formula = (
                f'=IF(F{current_row}<>"",F{current_row},"")'
                if key == "price"
                else (
                    f'=IF(F{current_row}<>"",F{current_row},IF(C{current_row}<>"",C{current_row},IF(B{current_row}<>"",B{current_row},IF(D{current_row}<>"",D{current_row},IF(E{current_row}<>"",E{current_row},0.25)))))'
                    if key == "scenario_tax_rate"
                    else (
                    _excel_manual_percent_active_formula(current_row)
                    if key == "fcf_yield"
                    else (
                    f'=IF(F{current_row}<>"",F{current_row},IF(D{current_row}<>"",D{current_row},IF(C{current_row}<>"",C{current_row},IF(B{current_row}<>"",B{current_row},E{current_row}))))'
                    if key == "capex"
                    else f'=IF(F{current_row}<>"",F{current_row},IF(C{current_row}<>"",C{current_row},IF(B{current_row}<>"",B{current_row},IF(D{current_row}<>"",D{current_row},E{current_row}))))'
                    )
                    )
                )
            )
            row = _write_cells(
                row,
                [label, fy_default, ttm_default, current_year_guidance, next_q_guidance, "", active_formula, note],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={1, 8},
                merge_spans=[(8, visible_max_col)],
            )
            for cc in range(2, 8):
                ws.cell(row=current_row, column=cc).number_format = number_format
                ws.cell(row=current_row, column=cc).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=current_row, column=6).fill = copy(input_fill)
            refs[key] = f"$G${current_row}"
            refs[f"{key}__latest_year"] = f"$B${current_row}"
            refs[f"{key}__ttm"] = f"$C${current_row}"
            refs[f"{key}__guidance_current"] = f"$D${current_row}"
            refs[f"{key}__guidance_next"] = f"$E${current_row}"
            refs[f"{key}__override"] = f"$F${current_row}"
        return row + 1, refs

    def _segment_scenario_specs(default_margin_proxy: Any = None) -> List[_SegmentScenarioInputSpec]:
        specs = _segment_scenario_specs_from_records(
            _records("Segment Scenario Inputs"),
            default_margin_proxy=default_margin_proxy,
            default_margin_basis="Active company operating margin proxy",
        )
        if specs:
            return specs
        fallback_rows = [
            ("Abercrombie (brand)", "Brand", _numeric_value("Brand Health", "Abercrombie 2025 sales")),
            ("Hollister (brand)", "Brand", _numeric_value("Brand Health", "Hollister 2025 sales")),
            ("Americas (geography / stores)", "Geography / stores", None),
            ("EMEA (geography / stores)", "Geography / stores", None),
            ("APAC (geography / stores)", "Geography / stores", None),
        ]
        out: List[_SegmentScenarioInputSpec] = []
        for label, category_type, raw_value in fallback_rows:
            out.append(
                _SegmentScenarioInputSpec(
                    label=label,
                    category_type=category_type,
                    baseline_revenue_m=_segment_scenario_revenue_m(raw_value, "$"),
                    revenue_basis="2025 year net sales" if raw_value is not None else "",
                    margin_conversion=default_margin_proxy if raw_value is not None and default_margin_proxy not in (None, "") else None,
                    margin_basis="Active company operating margin proxy" if raw_value is not None and default_margin_proxy not in (None, "") else "",
                    feeds_bridge=raw_value is not None and default_margin_proxy not in (None, ""),
                    source_note="Active company operating margin proxy" if raw_value is not None else "Missing segment revenue",
                    view_basis=_segment_scenario_view_basis(label, category_type),
                )
            )
        return out

    def _write_segment_scenario_inputs(row: int, refs: Dict[str, str]) -> Tuple[int, Dict[str, str]]:
        specs = _segment_scenario_specs(refs.get("operating_margin"))
        row = _section(row, "Segment Scenario Inputs")
        row = _headers(
            row,
            [
                "Segment / category",
                "Type",
                "Baseline revenue",
                "Revenue % change",
                "Revenue impact",
                "Operating margin",
                "EBITDA impact",
                "Feeds bridge?",
                "Notes",
            ],
            end_col=visible_max_col,
            merge_spans=[(9, visible_max_col)],
        )
        basis_row = row
        row = _write_cells(
            row,
            ["Active basis", "None", "", "", "", "", "", "", "Select category"],
            end_col=visible_max_col,
            fill=neutral,
            wrap_cols={1, 9},
            merge_spans=[(9, visible_max_col)],
        )
        ws.cell(row=basis_row, column=2).fill = copy(input_fill)
        validation = DataValidation(type="list", formula1='"None,Brand,Geography"', allow_blank=False)
        ws.add_data_validation(validation)
        validation.add(ws.cell(row=basis_row, column=2))

        def _subheader(title: str, row_num: int) -> int:
            sub_fill = PatternFill("solid", fgColor="DDEBF7")
            for cc in range(1, visible_max_col + 1):
                cell = ws.cell(row=row_num, column=cc)
                cell.fill = sub_fill
                cell.font = Font(bold=True, size=12, color=dark)
                cell.border = thin
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row=row_num, column=1, value=title)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=visible_max_col)
            ws.row_dimensions[row_num].height = 21
            return row_num + 1

        ordered_specs: List[_SegmentScenarioInputSpec] = []
        brand_specs = [spec for spec in specs if spec.view_basis == "Brand"]
        geo_specs = [spec for spec in specs if spec.view_basis == "Geography"]
        other_specs = [spec for spec in specs if spec.view_basis not in {"Brand", "Geography"}]
        if brand_specs:
            row = _subheader("Brand view — summed only when Brand selected", row)
            ordered_specs.extend(brand_specs)
        if geo_specs:
            geo_header_pending = True
        else:
            geo_header_pending = False
        for idx, spec in enumerate(brand_specs):
            current_row = row
            note = spec.source_note or ("Missing segment margin" if spec.margin_conversion is None else "Informational only")
            has_margin = spec.margin_conversion is not None
            ebitda_impact: Any = f'=IFERROR(IF(OR(D{current_row}="",F{current_row}=""),0,E{current_row}*F{current_row}),0)' if has_margin else ""
            feeds_formula: Any = (
                f'=IF(AND($B${basis_row}="Brand",C{current_row}<>"",D{current_row}<>"",F{current_row}<>""),"Yes","No")'
                if spec.feeds_bridge and has_margin and spec.baseline_revenue_m is not None
                else "No"
            )
            row = _write_cells(
                row,
                [
                    spec.label,
                    spec.category_type,
                    spec.baseline_revenue_m if spec.baseline_revenue_m is not None else "",
                    "",
                    f'=IFERROR(IF(D{current_row}="",0,C{current_row}*D{current_row}),0)',
                    spec.margin_conversion if spec.margin_conversion is not None else "",
                    ebitda_impact,
                    feeds_formula,
                    note,
                ],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={1, 2, 9},
                merge_spans=[(9, visible_max_col)],
            )
            sheet_ref = f"'{ws.title}'"
            spec.revenue_change_ref = f"={sheet_ref}!$D${current_row}"
            spec.revenue_impact_ref = f"={sheet_ref}!$E${current_row}"
            spec.ebitda_impact_ref = f"={sheet_ref}!$G${current_row}"
            spec.feeds_bridge_ref = f"={sheet_ref}!$H${current_row}"
            ws.cell(row=current_row, column=3).number_format = "#,##0.0"
            ws.cell(row=current_row, column=4).number_format = "0.0%"
            ws.cell(row=current_row, column=4).fill = copy(input_fill)
            ws.cell(row=current_row, column=5).number_format = "#,##0.0"
            ws.cell(row=current_row, column=6).number_format = "0.0%"
            ws.cell(row=current_row, column=7).number_format = "#,##0.0"
            ws.cell(row=current_row, column=8).alignment = Alignment(horizontal="left", vertical="center")
            for cc in range(1, visible_max_col + 1):
                ws.cell(row=current_row, column=cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc in {1, 2, 9})
        if geo_header_pending:
            row = _subheader("Geography view — summed only when Geography selected", row)
            ordered_specs.extend(geo_specs)
        for idx, spec in enumerate(geo_specs):
            current_row = row
            note = spec.source_note or ("Missing segment margin" if spec.margin_conversion is None else "Informational only")
            has_margin = spec.margin_conversion is not None
            ebitda_impact = f'=IFERROR(IF(OR(D{current_row}="",F{current_row}=""),0,E{current_row}*F{current_row}),0)' if has_margin else ""
            feeds_formula = (
                f'=IF(AND($B${basis_row}="Geography",C{current_row}<>"",D{current_row}<>"",F{current_row}<>""),"Yes","No")'
                if spec.feeds_bridge and has_margin and spec.baseline_revenue_m is not None
                else "No"
            )
            row = _write_cells(
                row,
                [
                    spec.label,
                    spec.category_type,
                    spec.baseline_revenue_m if spec.baseline_revenue_m is not None else "",
                    "",
                    f'=IFERROR(IF(D{current_row}="",0,C{current_row}*D{current_row}),0)',
                    spec.margin_conversion if spec.margin_conversion is not None else "",
                    ebitda_impact,
                    feeds_formula,
                    note,
                ],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={1, 2, 9},
                merge_spans=[(9, visible_max_col)],
            )
            sheet_ref = f"'{ws.title}'"
            spec.revenue_change_ref = f"={sheet_ref}!$D${current_row}"
            spec.revenue_impact_ref = f"={sheet_ref}!$E${current_row}"
            spec.ebitda_impact_ref = f"={sheet_ref}!$G${current_row}"
            spec.feeds_bridge_ref = f"={sheet_ref}!$H${current_row}"
            ws.cell(row=current_row, column=3).number_format = "#,##0.0"
            ws.cell(row=current_row, column=4).number_format = "0.0%"
            ws.cell(row=current_row, column=4).fill = copy(input_fill)
            ws.cell(row=current_row, column=5).number_format = "#,##0.0"
            ws.cell(row=current_row, column=6).number_format = "0.0%"
            ws.cell(row=current_row, column=7).number_format = "#,##0.0"
            for cc in range(1, visible_max_col + 1):
                ws.cell(row=current_row, column=cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc in {1, 2, 9})
        for idx, spec in enumerate(other_specs):
            current_row = row
            note = spec.source_note or ("Missing segment margin" if spec.margin_conversion is None else "Informational only")
            has_margin = spec.margin_conversion is not None
            ebitda_impact = f'=IFERROR(IF(OR(D{current_row}="",F{current_row}=""),0,E{current_row}*F{current_row}),0)' if has_margin else ""
            feeds_formula = (
                f'=IF(AND(C{current_row}<>"",D{current_row}<>"",F{current_row}<>""),"Yes","No")'
                if spec.feeds_bridge and has_margin and spec.baseline_revenue_m is not None
                else "No"
            )
            row = _write_cells(
                row,
                [
                    spec.label,
                    spec.category_type,
                    spec.baseline_revenue_m if spec.baseline_revenue_m is not None else "",
                    "",
                    f'=IFERROR(IF(D{current_row}="",0,C{current_row}*D{current_row}),0)',
                    spec.margin_conversion if spec.margin_conversion is not None else "",
                    ebitda_impact,
                    feeds_formula,
                    note,
                ],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={1, 2, 9},
                merge_spans=[(9, visible_max_col)],
            )
            sheet_ref = f"'{ws.title}'"
            spec.revenue_change_ref = f"={sheet_ref}!$D${current_row}"
            spec.revenue_impact_ref = f"={sheet_ref}!$E${current_row}"
            spec.ebitda_impact_ref = f"={sheet_ref}!$G${current_row}"
            spec.feeds_bridge_ref = f"={sheet_ref}!$H${current_row}"
            for cc in range(1, visible_max_col + 1):
                ws.cell(row=current_row, column=cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc in {1, 2, 9})
            ordered_specs.append(spec)
        _write_scenario_driver_assumptions_sheet(wb, ticker="ANF", segment_specs=ordered_specs, enabled=True)
        all_rows = [
            int(str(spec.ebitda_impact_ref).split("$G$")[-1])
            for spec in ordered_specs
            if "$G$" in str(spec.ebitda_impact_ref)
        ]
        selected_formula = (
            f'=IF($B${basis_row}="None",0,SUMIF(H{min(all_rows)}:H{max(all_rows)},"Yes",G{min(all_rows)}:G{max(all_rows)}))'
            if all_rows
            else ""
        )
        return row + 1, {"selected_segment_impact": selected_formula} if selected_formula else {}

    def _manual_ref(refs: Dict[str, str], key: str) -> str:
        return refs.get(key, '""')

    def _manual_part_ref(refs: Dict[str, str], key: str, part: str) -> str:
        return refs.get(f"{key}__{part}", '""')

    def _active_value_formula(ref: str) -> str:
        return f'=IF({ref}="","",{ref})'

    def _write_market_pricing(row: int, refs: Dict[str, str]) -> int:
        price = _manual_ref(refs, "price")
        shares = _manual_ref(refs, "shares")
        net_debt = _manual_ref(refs, "net_debt")
        eps = _manual_ref(refs, "eps")
        ebitda = _manual_ref(refs, "ebitda")
        fcf = _manual_ref(refs, "fcf")
        row = _section(row, "What Market Is Pricing")
        row = _headers(row, ["Metric", "Value / read", "", "", "Notes"], end_col=visible_max_col, merge_spans=[(2, 4), (5, visible_max_col)])
        market_rows = [
            ("Market price input", f'=IF({price}="","Manual current share price needed to calculate market-implied expectations.","Manual share price active; implied multiples use the active scenario inputs.")', "Leave current share price blank when no manual market price is intended.", "@"),
            ("Implied market cap", f'=IF({price}="","",{price}*{shares})', "Current share price x active share count.", "#,##0.0"),
            ("Implied EV", f'=IF(OR({price}="",{shares}="",{net_debt}=""),"",{price}*{shares}+{net_debt})', "Market cap plus net debt / less net cash.", "#,##0.0"),
            ("Implied P/E", f'=IFERROR(IF(OR({price}="",{eps}="",{eps}=0),"",{price}/{eps}),"")', "Price divided by active forward EPS.", "0.0x"),
            ("Implied EV/Adj EBITDA", f'=IFERROR(IF(OR({price}="",{shares}="",{ebitda}="",{ebitda}=0),"",({price}*{shares}+{net_debt})/{ebitda}),"")', "Implied enterprise value divided by active Adj EBITDA.", "0.0x"),
            ("Implied FCF yield", f'=IFERROR(IF(OR({price}="",{shares}="",{fcf}=""),"",{fcf}/({price}*{shares})),"")', "Active FCF divided by implied market cap.", "0.0%"),
        ]
        for idx, (metric, value, note, number_format) in enumerate(market_rows):
            current_row = row
            row = _write_cells(
                row,
                [metric, value, "", "", note],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={2, 5},
                merge_spans=[(2, 4), (5, visible_max_col)],
            )
            ws.cell(row=current_row, column=2).number_format = number_format
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return row

    def _write_scenario_driver_bridge(row: int, refs: Dict[str, str], segment_refs: Optional[Dict[str, str]] = None) -> Tuple[int, Dict[str, str]]:
        eps = _manual_ref(refs, "eps")
        ebitda = _manual_ref(refs, "ebitda")
        fcf = _manual_ref(refs, "fcf")
        shares = _manual_ref(refs, "shares")
        eps_override = _manual_part_ref(refs, "eps", "override")
        shares_ttm = _manual_part_ref(refs, "shares", "ttm")
        scenario_tax_rate = _manual_ref(refs, "scenario_tax_rate")
        segment_refs = segment_refs or {}
        selected_segment_impact = segment_refs.get("selected_segment_impact", "")
        refs_out: Dict[str, str] = {}
        row = _section(row, "Scenario Driver Bridge")
        row = _write_cells(
            row,
            ["Bridge-adjusted values start from active inputs and add incremental effects below | Taxable EPS impacts use active scenario tax rate."],
            end_col=visible_max_col,
            fill=neutral,
            merge_spans=[(1, visible_max_col)],
        )
        row = _headers(
            row,
            [
                "Bridge item",
                "Baseline included",
                "Active / guide",
                "Incremental effect",
                "EPS impact",
                "EBITDA impact",
                "FCF impact",
                "Read",
            ],
            end_col=visible_max_col,
            merge_spans=[(8, visible_max_col)],
        )
        incremental_start = row
        buyback_shares = _manual_ref(refs, "buyback_shares")
        capex = _manual_ref(refs, "capex")
        capex_baseline = _manual_part_ref(refs, "capex", "ttm")
        revenue = _manual_ref(refs, "revenue")
        margin_bps_refs = [
            _manual_ref(refs, "tariff_bps"),
            _manual_ref(refs, "freight_bps"),
            _manual_ref(refs, "erp_bps"),
            _manual_ref(refs, "marketing_bps"),
        ]
        margin_bridge_impact = (
            f'=IFERROR(IF({revenue}="",0,{revenue}*SUM('
            + ",".join(f'IF({ref}="",0,{ref})' for ref in margin_bps_refs)
            + ")/10000),0)"
        )
        bridge_specs = [
            _ScenarioDriverBridgeSpec(
                "Margin bridge vs baseline",
                SCENARIO_DRIVER_MARGIN_EBITDA,
                0,
                margin_bridge_impact,
                "Converts bps margin bridge to $m using active revenue.",
                explicit_incremental=True,
                ebitda_impact="same",
                fcf_impact="none",
                eps_impact="auto",
                tax_treatment=SCENARIO_TAX_TAXABLE,
                tax_source_basis="Tariff/freight/ERP/marketing bps rows converted to $m using active revenue.",
            ),
            _ScenarioDriverBridgeSpec(
                "Buyback/share-count effect",
                SCENARIO_DRIVER_SHARE_COUNT_BUYBACK,
                f'=IF({shares_ttm}="","Unknown",{shares_ttm})',
                _active_value_formula(buyback_shares),
                "Share count affects EPS.",
                reverse_incremental=True,
                ebitda_impact="none",
                fcf_impact="none",
                eps_impact="share_count",
                tax_treatment=SCENARIO_TAX_NO_EPS_IMPACT,
                tax_source_basis="Buybacks affect EPS through diluted shares, not earnings.",
            ),
            _ScenarioDriverBridgeSpec(
                "Selected segment revenue/margin impact",
                SCENARIO_DRIVER_MARGIN_EBITDA,
                0,
                selected_segment_impact or 0,
                "Selected Segment Scenario Inputs feed taxable EBITDA uplift.",
                explicit_incremental=True,
                ebitda_impact="same",
                fcf_impact="none",
                eps_impact="auto",
                tax_treatment=SCENARIO_TAX_TAXABLE,
                tax_source_basis="Segment Scenario Inputs selected Feeds bridge? rows.",
                audit_notes="ANF active basis prevents summing brand and geography together.",
            ),
            _ScenarioDriverBridgeSpec(
                "Capex change vs baseline",
                SCENARIO_DRIVER_CASH_FLOW_CAPEX,
                f'=IF({capex_baseline}="","Unknown",{capex_baseline})',
                _active_value_formula(capex),
                "Capex change affects FCF only.",
                ebitda_impact="none",
                fcf_impact="negative",
                eps_impact="none",
                tax_treatment=SCENARIO_TAX_CASH_ONLY,
                tax_source_basis="Capex affects cash flow, not direct EPS or Adj EBITDA.",
            ),
        ]
        after_tax_factor = None
        tax_source_basis = "Manual Market / Scenario Inputs active Scenario tax rate; defaults to 25% if no clean source."

        _write_scenario_bridge_tax_treatment_sheet(
            wb,
            ticker="ANF",
            specs=bridge_specs,
            after_tax_factor=after_tax_factor,
            tax_rate_ref=scenario_tax_rate,
            tax_source_basis=tax_source_basis,
        )
        incremental_rows = [
            _scenario_bridge_row_values(
                spec,
                row + idx,
                active_eps_ref=eps,
                active_shares_ref=shares,
                baseline_shares_ref=shares_ttm,
                eps_override_ref=eps_override,
                after_tax_factor=after_tax_factor,
                tax_rate_ref=scenario_tax_rate,
            )
            for idx, spec in enumerate(bridge_specs)
        ]
        for idx, (item, baseline, active, incremental, eps_impact, ebitda_impact, fcf_impact, read) in enumerate(incremental_rows):
            current_row = row
            row = _write_cells(
                row,
                [item, baseline, active, incremental, eps_impact, ebitda_impact, fcf_impact, read],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={8},
                merge_spans=[(8, visible_max_col)],
            )
            for cc in range(2, 8):
                ws.cell(row=current_row, column=cc).number_format = "#,##0.0"
                ws.cell(row=current_row, column=cc).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=current_row, column=5).number_format = "$0.00"
        incremental_end = row - 1
        row += 1
        row = _headers(
            row,
            ["Metric", "Active input", "", "Adjustment", "", "Bridge-adjusted value", "", "Read"],
            end_col=visible_max_col,
            merge_spans=[(2, 3), (4, 5), (6, 7), (8, visible_max_col)],
        )
        summary_start = row
        bridge_eps_value = _scenario_bridge_eps_value_formula(
            summary_row=summary_start,
            eps_override_ref=eps_override,
            active_eps_ref=eps,
            active_shares_ref=shares,
            baseline_shares_ref=shares_ttm,
            eps_impact_start=incremental_start,
            eps_impact_end=incremental_end,
        )
        eps_adjustment = f'=IFERROR(IF(B{summary_start}="","",{bridge_eps_value[1:]}-B{summary_start}),"")'
        summary_rows = [
            (
                "Bridge EPS ($/sh)",
                _active_value_formula(eps),
                eps_adjustment,
                bridge_eps_value,
                "Share count and bridge impacts apply.",
                "bridge_eps",
                "$0.00",
            ),
            (
                "Bridge Adj EBITDA ($m)",
                _active_value_formula(ebitda),
                f'=IFERROR(SUM(F{incremental_start}:F{incremental_end}),"")',
                f'=IFERROR(IF(B{summary_start + 1}="","",B{summary_start + 1}+D{summary_start + 1}),"")',
                "Active EBITDA plus incremental EBITDA impacts.",
                "bridge_ebitda",
                "#,##0.0",
            ),
            (
                "Bridge FCF ($m)",
                _active_value_formula(fcf),
                f'=IFERROR(SUM(G{incremental_start}:G{incremental_end}),"")',
                f'=IFERROR(IF(B{summary_start + 2}="","",B{summary_start + 2}+D{summary_start + 2}),"")',
                "Active FCF plus incremental FCF impacts.",
                "bridge_fcf",
                "#,##0.0",
            ),
        ]
        for idx, (metric, active, adjustment, bridge_value, read, ref_key, number_format) in enumerate(summary_rows):
            current_row = row
            row = _write_cells(
                row,
                [metric, active, "", adjustment, "", bridge_value, "", read],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={8},
                merge_spans=[(2, 3), (4, 5), (6, 7), (8, visible_max_col)],
            )
            for cc in (2, 4, 6):
                ws.cell(row=current_row, column=cc).number_format = number_format
                ws.cell(row=current_row, column=cc).alignment = Alignment(horizontal="left", vertical="center")
            if ref_key:
                refs_out[ref_key] = f"$F${current_row}"
        return row + 1, refs_out

    def _write_manual_scenarios(row: int, refs: Dict[str, str], bridge_refs: Optional[Dict[str, str]] = None) -> int:
        bridge_refs = bridge_refs or {}
        eps = bridge_refs.get("bridge_eps") or _manual_ref(refs, "eps")
        ebitda = bridge_refs.get("bridge_ebitda") or _manual_ref(refs, "ebitda")
        fcf = bridge_refs.get("bridge_fcf") or _manual_ref(refs, "fcf")
        pe = _manual_ref(refs, "pe")
        ev_multiple = _manual_ref(refs, "ev_multiple")
        fcf_yield = _manual_ref(refs, "fcf_yield")
        fcf_yield_rate = _excel_percent_value_expr(fcf_yield)
        shares = _manual_ref(refs, "shares")
        net_debt = _manual_ref(refs, "net_debt")
        row = _section(row, "Bear / Base / Bull Scenario")
        row = _headers(
            row,
            ["Scenario", "Key assumptions", "", "EPS", "Adj EBITDA", "FCF", "Value/share @ P/E", "Value/share @ EV/Adj EBITDA", "Value/share @ FCF yield", "Value range"],
            end_col=visible_max_col,
            merge_spans=[(2, 3)],
        )
        scenarios = [
            ("Bear", "Lower sales/margin, lower multiple and higher FCF yield.", 0.90, 0.90, 0.85, 0.90, 0.90, 1.15),
            ("Base", "Active manual/default assumptions.", 1.00, 1.00, 1.00, 1.00, 1.00, 1.00),
            ("Bull", "Better sales/margin, stronger multiple and lower FCF yield.", 1.10, 1.10, 1.15, 1.15, 1.15, 0.85),
        ]
        for idx, (name, assumptions, eps_factor, ebitda_factor, fcf_factor, pe_factor, ev_factor, yield_factor) in enumerate(scenarios):
            current_row = row
            eps_cell = f"D{current_row}"
            ebitda_cell = f"E{current_row}"
            fcf_cell = f"F{current_row}"
            row = _write_cells(
                row,
                [
                    name,
                    assumptions,
                    "",
                    f'=IFERROR(IF({eps}="","",{eps}*{eps_factor}),"")',
                    f'=IFERROR(IF({ebitda}="","",{ebitda}*{ebitda_factor}),"")',
                    f'=IFERROR(IF({fcf}="","",{fcf}*{fcf_factor}),"")',
                    f'=IFERROR(IF(OR({eps_cell}="",{eps_cell}<=0,{pe}=""),"N/M",{eps_cell}*{pe}*{pe_factor}),"N/M")',
                    f'=IFERROR(IF(OR({ebitda_cell}="",{ev_multiple}="",{shares}="",{shares}=0),"",({ebitda_cell}*{ev_multiple}*{ev_factor}-{net_debt})/{shares}),"")',
                    f'=IFERROR(IF(OR({fcf_cell}="",{fcf_yield}="",{shares}="",{shares}=0),"",({fcf_cell}/(({fcf_yield_rate})*{yield_factor}))/{shares}),"")',
                    _excel_visible_value_range_formula(current_row),
                ],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={2},
                merge_spans=[(2, 3)],
            )
            for cc in (4, 7, 8, 9):
                ws.cell(row=current_row, column=cc).number_format = "$0.00"
            for cc in (5, 6):
                ws.cell(row=current_row, column=cc).number_format = "#,##0.0"
            ws.row_dimensions[current_row].height = 30
        row = _write_cells(
            row,
            ["Uses Investment_Case manual inputs; may differ from Valuation Thesis Bridge."],
            end_col=visible_max_col,
            fill=callout_fill,
            wrap_cols={1},
            merge_spans=[(1, visible_max_col)],
        )
        return row

    def _write_manual_valuation_sensitivity(row: int, refs: Dict[str, str], bridge_refs: Optional[Dict[str, str]] = None) -> int:
        bridge_refs = bridge_refs or {}
        eps = bridge_refs.get("bridge_eps") or _manual_ref(refs, "eps")
        ebitda = bridge_refs.get("bridge_ebitda") or _manual_ref(refs, "ebitda")
        fcf = bridge_refs.get("bridge_fcf") or _manual_ref(refs, "fcf")
        pe = _manual_ref(refs, "pe")
        ev_multiple = _manual_ref(refs, "ev_multiple")
        fcf_yield = _manual_ref(refs, "fcf_yield")
        fcf_yield_rate = _excel_percent_value_expr(fcf_yield)
        shares = _manual_ref(refs, "shares")
        net_debt = _manual_ref(refs, "net_debt")
        row = _section(row, "Valuation Sensitivity")
        row = _headers(row, ["EPS", "10x", "12x", "14x", "16x"], end_col=visible_max_col)
        for idx, factor in enumerate((0.90, 1.00, 1.10)):
            current_row = row
            row = _write_cells(
                row,
                [
                    f'=IFERROR(IF({eps}="","",{eps}*{factor}),"")',
                    f"=$A{current_row}*10",
                    f"=$A{current_row}*12",
                    f"=$A{current_row}*14",
                    f"=$A{current_row}*16",
                ],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
            )
            for cc in range(1, 6):
                ws.cell(row=current_row, column=cc).number_format = "$0.00"
                ws.cell(row=current_row, column=cc).alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        row = _headers(row, ["Scenario", "EPS", "P/E", "Share price"], end_col=visible_max_col)
        for idx, (name, eps_factor, pe_factor) in enumerate((("Bear", 0.90, 0.90), ("Base", 1.00, 1.00), ("Bull", 1.10, 1.15))):
            current_row = row
            row = _write_cells(
                row,
                [
                    name,
                    f'=IFERROR(IF({eps}="","",{eps}*{eps_factor}),"")',
                    f'=IFERROR({pe}*{pe_factor},"")',
                    f'=IFERROR(B{current_row}*C{current_row},"")',
                ],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
            )
            ws.cell(row=current_row, column=2).number_format = "$0.00"
            ws.cell(row=current_row, column=3).number_format = "0.0x"
            ws.cell(row=current_row, column=4).number_format = "$0.00"
        row += 1

        row = _section(row, "Adj EBITDA x EV/EBITDA")
        row = _headers(row, ["Multiple", "EV", "Equity value: core net cash", "Share price", "Source / investment read"], end_col=visible_max_col, merge_spans=[(5, visible_max_col)])
        for idx, delta in enumerate((-2.0, 0.0, 2.0)):
            current_row = row
            row = _write_cells(
                row,
                [
                    f'=IFERROR(MAX(0,{ev_multiple}+{delta}),"")',
                    f'=IFERROR(A{current_row}*{ebitda},"")',
                    f'=IFERROR(B{current_row}-{net_debt},"")',
                    f'=IFERROR(C{current_row}/{shares},"")',
                    "Active Adj EBITDA x scenario EV/EBITDA, less active net debt.",
                ],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={5},
                merge_spans=[(5, visible_max_col)],
            )
            ws.cell(row=current_row, column=1).number_format = "0.0x"
            for cc in (2, 3):
                ws.cell(row=current_row, column=cc).number_format = "#,##0.0"
            ws.cell(row=current_row, column=4).number_format = "$0.00"
        row += 1

        row = _section(row, "FCF Yield Implied Equity Value")
        row = _headers(row, ["Yield", "Equity value", "Share price", "Source / note"], end_col=visible_max_col, merge_spans=[(4, visible_max_col)])
        for idx, factor in enumerate((0.80, 1.00, 1.20)):
            current_row = row
            row = _write_cells(
                row,
                [
                    f'=IFERROR(({fcf_yield_rate})*{factor},"")',
                    f'=IFERROR({fcf}/A{current_row},"")',
                    f'=IFERROR(B{current_row}/{shares},"")',
                    "Active FCF capitalized by scenario equity FCF yield.",
                ],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={4},
                merge_spans=[(4, visible_max_col)],
            )
            ws.cell(row=current_row, column=1).number_format = "0.0%"
            ws.cell(row=current_row, column=2).number_format = "#,##0.0"
            ws.cell(row=current_row, column=3).number_format = "$0.00"
        return row

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=visible_max_col)
    ws.cell(row=1, column=1, value="ANF Investment Case")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(1, visible_max_col + 1):
        ws.cell(row=1, column=cc).fill = title_fill
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=visible_max_col)
    ws.cell(row=2, column=1, value="Quarter labels are fiscal periods; for ANF, 2025-Q4 ended 2026-01-31. Market price is not required; sensitivity tables are scenario based.")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color=muted)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")
    row = 4

    row = _section(row, "Investment Snapshot")
    snapshot_rows = [
        ("Model read", _display("Investment Snapshot", "Model read")),
        ("Why it can work", _display("Investment Snapshot", "Why it can work")),
        ("Key debate", _display("Investment Snapshot", "Key debate")),
        ("Upside path", _display("Investment Snapshot", "Upside path") or _display("Investment Snapshot", "What improves case")),
        ("Downside path", _display("Investment Snapshot", "Downside path") or _display("Investment Snapshot", "What breaks case")),
        ("Watch next", _display("Investment Snapshot", "Watch next")),
        ("Current stance", _display("Investment Snapshot", "Current stance based on model data")),
    ]
    for idx, vals in enumerate(snapshot_rows):
        row = _write_cells(
            row,
            [vals[0], vals[1]],
            end_col=visible_max_col,
            fill=callout_fill if idx == 0 else (neutral_alt if idx % 2 == 1 else neutral),
            wrap_cols={2},
            merge_spans=[(2, visible_max_col)],
        )
        ws.cell(row=row - 1, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row - 1, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    row += 1

    row, manual_refs = _write_manual_inputs(row)
    row, segment_refs = _write_segment_scenario_inputs(row, manual_refs)
    row, bridge_refs = _write_scenario_driver_bridge(row, manual_refs, segment_refs)
    row = _write_market_pricing(row, manual_refs)
    row += 1

    row = _section(row, "Key Debate")
    row = _write_cells(row, ["Debate", _display("Key Debate", "Key debate")], end_col=visible_max_col, fill=callout_fill, wrap_cols={2}, merge_spans=[(2, visible_max_col)])
    ws.cell(row=row - 1, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=row - 1, column=2).font = Font(bold=True, size=12, color=dark)
    ws.row_dimensions[row - 1].height = 21
    row += 1

    key_debate_rows = _records("Key Debates")
    if key_debate_rows:
        row = _section(row, "Key Debates")
        for idx, rec in enumerate(key_debate_rows):
            title = str(rec.get("metric", "") or "").strip()
            current_read = str(rec.get("current_read", "") or "").strip()
            header_text = title
            if current_read:
                header_text = f"{title} | Current read: {current_read}" if title else f"Current read: {current_read}"
            row = _write_cells(
                row,
                [header_text],
                end_col=visible_max_col,
                fill=subheader_fill,
                wrap_cols={1},
                merge_spans=[(1, visible_max_col)],
            )
            for cc in range(1, visible_max_col + 1):
                ws.cell(row=row - 1, column=cc).font = Font(bold=True, size=12, color=dark)
                ws.cell(row=row - 1, column=cc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row - 1].height = 24
            row = _write_cells(
                row,
                [
                    "Bull evidence",
                    rec.get("bull_evidence", "") or rec.get("display", ""),
                    "",
                    "",
                    "",
                    "Bear evidence",
                    rec.get("bear_evidence", ""),
                ],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={2, 7},
                merge_spans=[(2, 5), (7, visible_max_col)],
            )
            ws.cell(row=row - 1, column=1).font = Font(bold=True, size=12, color=dark)
            ws.cell(row=row - 1, column=6).font = Font(bold=True, size=12, color=dark)
            for cc in range(1, visible_max_col + 1):
                ws.cell(row=row - 1, column=cc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row - 1].height = 38
            row = _write_cells(
                row,
                ["Next proof point", rec.get("next_proof_point", "")],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={2},
                merge_spans=[(2, visible_max_col)],
            )
            ws.cell(row=row - 1, column=1).font = Font(bold=True, size=12, color=dark)
            for cc in range(1, visible_max_col + 1):
                ws.cell(row=row - 1, column=cc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row - 1].height = 26
        row += 1

    row = _write_manual_scenarios(row, manual_refs, bridge_refs)
    row += 1

    quality_rows = _records("Quality of Earnings")
    if quality_rows:
        row = _section(row, "Quality of Earnings")
        row = _headers(
            row,
            ["Item", "Impact", "", "Cash?", "Recurring?", "Read"],
            end_col=visible_max_col,
            merge_spans=[(2, 3), (6, visible_max_col)],
        )
        for idx, rec in enumerate(quality_rows):
            row = _write_cells(
                row,
                [
                    rec.get("metric", ""),
                    rec.get("display", ""),
                    "",
                    rec.get("cash_flag", ""),
                    rec.get("recurring_flag", ""),
                    rec.get("quality_read", ""),
                ],
                end_col=visible_max_col,
                fill=neutral_alt if idx % 2 == 0 else neutral,
                wrap_cols={2, 6},
                merge_spans=[(2, 3), (6, visible_max_col)],
            )
            ws.row_dimensions[row - 1].height = 24
        row += 1

    row = _section(row, "What Needs To Happen For The Stock To Work")
    row = _headers(row, ["Condition", "Current read"], end_col=visible_max_col, merge_spans=[(2, visible_max_col)])
    for idx, rec in enumerate(_records("What Needs To Happen")):
        row = _write_cells(
            row,
            [rec.get("metric", ""), rec.get("display", "")],
            end_col=visible_max_col,
            fill=neutral_alt if idx % 2 == 0 else neutral,
            wrap_cols={2},
            merge_spans=[(2, visible_max_col)],
        )
    row += 1

    row = _section(row, "Tariff / Margin Bridge", end_col=overflow_max_col)
    bridge_spans = [(2, 3), (4, 5), (6, overflow_max_col)]
    row = _headers(row, ["Bridge item", "Q1 2026", "", "2026 year", "", "Comment"], end_col=overflow_max_col, merge_spans=bridge_spans)
    summary_bridge_rows = [
        ("2025 op margin", "", _display("Tariff / Margin Bridge", "Reported 2025 operating margin"), "reported actual"),
        ("2026 op margin guide", "", _display("Tariff / Margin Bridge", "2026 guide operating margin"), "company guide"),
        ("Implied decline", "", _display("Tariff / Margin Bridge", "Implied decline"), "margin normalization embedded in guide"),
    ]
    for idx, vals in enumerate(summary_bridge_rows):
        row = _write_cells(
            row,
            [vals[0], vals[1], "", vals[2], "", vals[3]],
            end_col=overflow_max_col,
            fill=callout_fill if idx == 0 else (neutral_alt if idx % 2 == 0 else neutral),
            wrap_cols={6},
            merge_spans=bridge_spans,
        )
    bridge_rows = [
        ("Tariff headwind", _rec("Tariff / Margin Bridge", "Q1 2026 tariff headwind").get("q1_display", ""), _rec("Tariff / Margin Bridge", "2026 tariff headwind").get("year_display", ""), "core pressure"),
        ("Freight tailwind", _rec("Tariff / Margin Bridge", "Freight tailwind").get("q1_display", ""), _rec("Tariff / Margin Bridge", "Freight tailwind").get("year_display", ""), ""),
        ("ERP disruption", _rec("Tariff / Margin Bridge", "ERP disruption").get("q1_display", ""), _rec("Tariff / Margin Bridge", "ERP disruption").get("year_display", ""), ""),
        ("Marketing", _rec("Tariff / Margin Bridge", "Marketing").get("q1_display", ""), _rec("Tariff / Margin Bridge", "Marketing").get("year_display", ""), ""),
        ("AUR / pricing", _rec("Tariff / Margin Bridge", "AUR / pricing").get("q1_display", ""), _rec("Tariff / Margin Bridge", "AUR / pricing").get("year_display", ""), ""),
        ("Sourcing / supplier mitigation", _rec("Tariff / Margin Bridge", "Sourcing / supplier mitigation").get("q1_display", ""), _rec("Tariff / Margin Bridge", "Sourcing / supplier mitigation").get("year_display", ""), ""),
    ]
    for idx, vals in enumerate(bridge_rows):
        row = _write_cells(
            row,
            [vals[0], vals[1], "", vals[2], "", vals[3]],
            end_col=overflow_max_col,
            fill=neutral_alt if idx % 2 == 0 else neutral,
            wrap_cols={6},
            merge_spans=bridge_spans,
        )
    row = _write_cells(
        row,
        [
            "Bridge read",
            "",
            "",
            "",
            "",
            _display("Tariff / Margin Bridge", "Bridge read"),
        ],
        end_col=overflow_max_col,
        fill=callout_fill,
        wrap_cols={6},
        merge_spans=bridge_spans,
    )
    row += 1

    row = _section(row, "EPS Bridge")
    eps_spans = [(2, 5), (6, visible_max_col)]
    row = _headers(row, ["Bridge item", "Value / direction", "", "", "", "Investment read"], end_col=visible_max_col, merge_spans=eps_spans)
    eps_items = [
        ("2025 adjusted EPS", _display("EPS Bridge", "2025 adjusted EPS"), "clean starting point"),
        ("Sales growth", _display("EPS Bridge", "Sales growth"), "positive if guide holds"),
        ("Margin / tariff / freight / AUR", _display("EPS Bridge", "Margin / tariff / freight / AUR"), "key debate"),
        ("SG&A leverage / deleverage", _display("EPS Bridge", "SG&A leverage / deleverage"), "depends demand and marketing"),
        ("Buyback / share count reduction", _display("EPS Bridge", "Buyback / share count reduction"), "supports EPS"),
        ("2026 guided EPS", _display("EPS Bridge", "2026 guided EPS"), "current guide range"),
    ]
    for idx, vals in enumerate(eps_items):
        row = _write_cells(
            row,
            [vals[0], vals[1], "", "", "", vals[2]],
            end_col=visible_max_col,
            fill=neutral_alt if idx % 2 == 0 else neutral,
            wrap_cols={2, 6},
            merge_spans=eps_spans,
        )
    row += 1

    row = _section(row, "Buybacks vs FCF")
    row = _headers(row, ["Metric", "Latest", "Investment read"], end_col=visible_max_col, merge_spans=[(3, visible_max_col)])
    for idx, rec in enumerate(_records("Buybacks vs FCF")):
        is_read_row = str(rec.get("metric") or "") == "Investment read"
        row = _write_cells(
            row,
            [
                rec.get("metric", ""),
                "" if is_read_row else rec.get("display", ""),
                rec.get("display", "") if is_read_row else (rec.get("source_note", "") or rec.get("source", "")),
            ],
            end_col=visible_max_col,
            fill=callout_fill if is_read_row else (neutral_alt if idx % 2 == 0 else neutral),
            wrap_cols={3},
            merge_spans=[(3, visible_max_col)],
        )
    row += 1

    guide_section = "2026 Guide → Implied Earnings"
    row = _section(row, guide_section)
    guide_spans = [(2, 3), (4, visible_max_col)]
    row = _headers(row, ["Bridge line", "Value", "", "Source / note"], end_col=visible_max_col, merge_spans=guide_spans)
    guide_metrics = [
        "2025 revenue",
        "2026 revenue growth guide",
        "Implied 2026 revenue",
        "Operating margin guide",
        "Implied EBIT",
        "Tax / interest assumptions",
        "Diluted shares guide",
        "Implied EPS low/high",
        "Company EPS guide",
        "Model vs guide check",
    ]
    for idx, metric in enumerate(guide_metrics):
        rec = _rec(guide_section, metric)
        row = _write_cells(
            row,
            [metric, rec.get("display", ""), "", rec.get("source_note", "") or rec.get("source", "")],
            end_col=visible_max_col,
            fill=neutral_alt if idx % 2 == 0 else neutral,
            wrap_cols={4},
            merge_spans=guide_spans,
        )
    row += 1

    row = _section(row, "What Moves EPS", end_col=overflow_max_col)
    moves_spans = [(2, 5), (6, overflow_max_col)]
    row = _headers(row, ["Sensitivity", "Approx EPS impact", "", "", "", "Investment read"], end_col=overflow_max_col, merge_spans=moves_spans)
    for idx, metric in enumerate(["+100 bps operating margin", "+100 bps gross margin", "+1% sales growth", "$100m buybacks", "Roughly +$1 EPS equals"]):
        rec = _rec("What Moves EPS", metric)
        row = _write_cells(row, [metric, rec.get("display", ""), "", "", "", rec.get("source_note", "") or rec.get("source", "")], end_col=overflow_max_col, fill=neutral_alt if idx % 2 == 0 else neutral, wrap_cols={2, 6}, merge_spans=moves_spans)
    row += 1

    row = _write_manual_valuation_sensitivity(row, manual_refs, bridge_refs)
    row += 1

    row = _section(row, "Comp Stack / Lapping Risk", end_col=overflow_max_col)
    row = _headers(row, ["Quarter", "Total comp", "2Y stack", "Abercrombie comp", "Hollister comp", "Americas", "EMEA", "APAC", "Read"], end_col=overflow_max_col, merge_spans=[(9, overflow_max_col)])
    comp_recs = _records("Comp Stack / Lapping Risk")[-8:]
    for idx, rec in enumerate(comp_recs):
        row = _write_cells(
            row,
            [
                rec.get("quarter_label", ""),
                rec.get("total_comp", ""),
                rec.get("two_year_stack", ""),
                rec.get("abercrombie_comp", ""),
                rec.get("hollister_comp", ""),
                rec.get("americas", ""),
                rec.get("emea", ""),
                rec.get("apac", ""),
                rec.get("short_read", ""),
            ],
            end_col=overflow_max_col,
            fill=neutral_alt if idx % 2 == 0 else neutral,
            wrap_cols={9},
            merge_spans=[(9, overflow_max_col)],
        )
    row += 1

    row = _section(row, "Brand Health")
    row = _headers(row, ["Metric", "Abercrombie", "Hollister"], end_col=visible_max_col)
    brand_rows = [
        ("2025 sales", _display("Brand Health", "Abercrombie 2025 sales"), _display("Brand Health", "Hollister 2025 sales")),
        ("2025 sales growth", _display("Brand Health", "Abercrombie 2025 sales growth"), _display("Brand Health", "Hollister 2025 sales growth")),
        ("Q4 sales growth", _display("Brand Health", "Abercrombie Q4 sales growth"), _display("Brand Health", "Hollister Q4 sales growth")),
        ("Q4 comp", _display("Brand Health", "Abercrombie Q4 comp"), _display("Brand Health", "Hollister Q4 comp")),
    ]
    for idx, vals in enumerate(brand_rows):
        row = _write_cells(row, vals, end_col=visible_max_col, fill=neutral_alt if idx % 2 == 0 else neutral)
    row = _write_cells(row, ["Interpretation", _display("Brand Health", "Interpretation"), ""], end_col=visible_max_col, fill=callout_fill, wrap_cols={2}, merge_spans=[(2, visible_max_col)])
    row += 1

    row = _section(row, "Inventory / Markdown Risk")
    row = _headers(row, ["Metric", "Latest", "Interpretation"], end_col=visible_max_col, merge_spans=[(3, visible_max_col)])
    inv_metrics = ["Inventory growth", "Sales growth", "Inventory cost tariff component", "Inventory unit growth", "ERP prebuild component", "Ex-ERP unit growth"]
    for idx, metric in enumerate(inv_metrics):
        row = _write_cells(row, [metric, _display("Inventory / Markdown Risk", metric), ""], end_col=visible_max_col, fill=neutral_alt if idx % 2 == 0 else neutral, merge_spans=[(3, visible_max_col)])
    row = _write_cells(row, ["Conclusion", _display("Inventory / Markdown Risk", "Conclusion"), ""], end_col=visible_max_col, fill=callout_fill, wrap_cols={2}, merge_spans=[(2, visible_max_col)])
    row += 1

    row = _section(row, "Store Productivity / Real Estate ROI")
    store_spans = [(2, 5), (6, visible_max_col)]
    row = _headers(row, ["Metric", "2025 actual / 2026 guide", "", "", "", "Read"], end_col=visible_max_col, merge_spans=store_spans)
    store_metrics = ["Company-owned stores", "Franchise stores", "Total incl franchise", "2025 openings", "2025 closures", "2026 openings", "2026 closures", "2026 remodels/right-sizes", "Sales per owned store", "Method note", "Store growth", "Revenue growth vs store growth", "Digital mix"]
    for idx, metric in enumerate(store_metrics):
        rec = _rec("Store Productivity / Real Estate ROI", metric)
        row = _write_cells(row, [metric, _display("Store Productivity / Real Estate ROI", metric), "", "", "", rec.get("source_note", "") or rec.get("source", "")], end_col=visible_max_col, fill=neutral_alt if idx % 2 == 0 else neutral, wrap_cols={2, 6}, merge_spans=store_spans)
    row += 1

    row = _section(row, "Guidance Beat/Miss Setup")
    beat_spans = [(2, 3), (4, 7), (8, visible_max_col)]
    row = _headers(row, ["Metric", "Guide", "", "Current trend", "", "", "", "Beat/miss risk"], end_col=visible_max_col, merge_spans=beat_spans)
    for idx, rec in enumerate(_records("Guidance Beat/Miss Setup")):
        row = _write_cells(row, [rec.get("metric", ""), rec.get("display", ""), "", rec.get("current_trend", ""), "", "", "", rec.get("beat_miss_risk", "")], end_col=visible_max_col, fill=neutral_alt if idx % 2 == 0 else neutral, wrap_cols={4, 8}, merge_spans=beat_spans)

    ws.freeze_panes = "A4"
    widths = {
        "A": 42,
        "B": 32,
        "C": 30,
        "D": 28,
        "E": 32,
        "F": 24,
        "G": 24,
        "H": 24,
        "I": 22,
        "J": 22,
        "K": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row_cells in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=overflow_max_col):
        for cell in row_cells:
            if cell.value in (None, ""):
                continue
            cell.alignment = Alignment(
                horizontal="left",
                vertical=cell.alignment.vertical or "center",
                wrap_text=bool(cell.alignment.wrap_text),
            )
    for rr in range(1, ws.max_row + 1):
        if ws.row_dimensions[rr].height is None:
            ws.row_dimensions[rr].height = 19
