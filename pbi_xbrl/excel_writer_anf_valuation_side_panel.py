"""ANF Valuation side-panel support and rendering."""
from __future__ import annotations

from copy import copy as _copy
from dataclasses import dataclass
from typing import Any, Dict, MutableMapping, Optional, Sequence, Set, Tuple

from openpyxl.comments import Comment as _Comment
from openpyxl.styles import Alignment as _Alignment, Border as _Border, PatternFill as _PatternFill, Side as _Side
from openpyxl.utils import get_column_letter as _get_column_letter


@dataclass(frozen=True)
class AnfValuationSidePanelDeps:
    runtime: MutableMapping[str, Any]


def clear_anf_valuation_side_panels(
    deps: AnfValuationSidePanelDeps,
    ws: Any,
    *,
    start_col: int = 15,
    end_col: Optional[int] = None,
    side_max_row: int = 125,
) -> None:
    """Remove ANF-only narrative side panels from Valuation without touching the core model grid."""
    runtime = deps.runtime
    PatternFill = runtime.get("PatternFill", _PatternFill)
    Border = runtime.get("Border", _Border)
    Alignment = runtime.get("Alignment", _Alignment)
    if ws is None:
        return
    c1 = int(start_col)
    c2 = int(end_col or getattr(ws, "max_column", c1))
    side_tokens = (
        "Guidance (As of",
        "No guidance items for this quarter",
        "Operating Drivers",
        "Thesis Bridge",
        "Hidden Value Panel",
    )
    main_tokens = (
        "Guidance detail",
        "Current guidance",
        "See Promise_Progress_UI",
    )

    def _row_text(row_idx: int, first_col: int, last_col: int) -> str:
        return "\n".join(str(ws.cell(row_idx, cc).value or "") for cc in range(first_col, last_col + 1))

    def _side_row_has_token(row_idx: int) -> bool:
        txt = _row_text(row_idx, c1, c2)
        return any(tok in txt for tok in side_tokens)

    def _unmerge_overlaps(r1: int, r2: int, first_col: int, last_col: int) -> None:
        try:
            ranges = list(ws.merged_cells.ranges)
        except Exception:
            ranges = []
        for merged in ranges:
            try:
                min_col, min_row, max_col, max_row = merged.bounds
            except Exception:
                continue
            if max_row < r1 or min_row > r2 or max_col < first_col or min_col > last_col:
                continue
            try:
                ws.unmerge_cells(str(merged))
            except Exception:
                pass

    def _clear_rect(r1: int, r2: int, first_col: int, last_col: int) -> None:
        _unmerge_overlaps(r1, r2, first_col, last_col)
        for rr in range(r1, r2 + 1):
            for cc in range(first_col, last_col + 1):
                cell = ws.cell(rr, cc)
                cell.value = None
                cell.comment = None
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
                cell.alignment = Alignment()

    max_sheet_row = int(getattr(ws, "max_row", 0) or 0)
    side_scan_max_row = min(max_sheet_row, int(side_max_row or max_sheet_row))
    rr = 1
    while rr <= side_scan_max_row:
        if not _side_row_has_token(rr):
            rr += 1
            continue
        block_end = rr
        blank_streak = 0
        scan_limit = min(side_scan_max_row, rr + 45)
        for cand in range(rr + 1, scan_limit + 1):
            if _side_row_has_token(cand):
                break
            has_side_value = any(ws.cell(cand, cc).value not in (None, "") for cc in range(c1, c2 + 1))
            if has_side_value:
                block_end = cand
                blank_streak = 0
            else:
                blank_streak += 1
                if blank_streak >= 2:
                    break
        _clear_rect(rr, block_end, c1, c2)
        rr = block_end + 1

    for row_idx in range(1, max_sheet_row + 1):
        txt = _row_text(row_idx, 1, min(c2, int(getattr(ws, "max_column", c2) or c2)))
        if any(tok in txt for tok in main_tokens):
            _clear_rect(row_idx, row_idx, 1, min(c2, int(getattr(ws, "max_column", c2) or c2)))


def valuation_side_panel_style_bundle(deps: AnfValuationSidePanelDeps) -> Dict[str, Any]:
    return deps.runtime["_style_valuation_side_panel_style_bundle"]()


def write_anf_valuation_side_panel(
    deps: AnfValuationSidePanelDeps,
    ws: Any,
    *,
    start_row: int = 7,
    start_col: int = 15,
    end_col: int = 29,
) -> Dict[str, int]:
    """Render the intentional ANF valuation side panel after stray panels have been scrubbed."""
    runtime = deps.runtime
    PatternFill = runtime.get("PatternFill", _PatternFill)
    Border = runtime.get("Border", _Border)
    Side = runtime.get("Side", _Side)
    Alignment = runtime.get("Alignment", _Alignment)
    Comment = runtime.get("Comment", _Comment)
    get_column_letter = runtime.get("get_column_letter", _get_column_letter)
    copy = runtime.get("copy", _copy)
    _anf_clean_visible_ui_text = runtime["_anf_clean_visible_ui_text"]
    _shared_visible_period_text = runtime["_shared_visible_period_text"]
    if ws is None:
        return {}
    c1 = int(start_col)
    c2 = int(end_col)
    if c2 < c1:
        c1, c2 = c2, c1
    panel_width = c2 - c1 + 1
    if panel_width < 8:
        c2 = c1 + 7
    row = int(start_row)
    style = valuation_side_panel_style_bundle(deps)
    section_fill = copy(style["section_fill"])
    header_fill = copy(style["header_fill"])
    neutral_fill = copy(style["neutral_fill"])
    neutral_alt_fill = copy(style["neutral_alt_fill"])
    input_fill = copy(style["input_fill"])
    thin_border = copy(style["thin_border"])
    spacer_border = Border(
        top=Side(style="thin", color="D9E2EA"),
        bottom=Side(style="thin", color="D9E2EA"),
    )
    title_font = copy(style["title_font"])
    header_font = copy(style["header_font"])
    body_font = copy(style["body_font"])
    input_font = copy(style["input_font"])

    def _set_row_height_min(row_idx: int, height: float) -> None:
        """Never shrink Valuation rows: the main quarterly grid shares them."""
        current = ws.row_dimensions[row_idx].height
        if current is None or float(current) < float(height):
            ws.row_dimensions[row_idx].height = float(height)

    def _safe_text(value: Any, max_chars: int = 160) -> str:
        txt = _anf_clean_visible_ui_text(value, max_chars=max_chars)
        return _shared_visible_period_text(txt)

    def _unmerge_overlaps(r1: int, r2: int, first_col: int, last_col: int) -> None:
        for merged in list(getattr(ws, "merged_cells", ()).ranges):
            try:
                min_col, min_row, max_col, max_row = merged.bounds
            except Exception:
                continue
            if max_row < r1 or min_row > r2 or max_col < first_col or min_col > last_col:
                continue
            try:
                ws.unmerge_cells(str(merged))
            except Exception:
                pass

    def _clear_panel_rect(r1: int, r2: int, first_col: int, last_col: int) -> None:
        _unmerge_overlaps(r1, r2, first_col, last_col)
        for rr in range(r1, r2 + 1):
            for cc in range(first_col, last_col + 1):
                cell = ws.cell(rr, cc)
                cell.value = None
                cell.comment = None
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
                cell.alignment = Alignment()

    def _comment(cell: Any, text: str) -> None:
        txt = _safe_text(text, max_chars=1000)
        if not txt:
            return
        try:
            cell.comment = Comment(txt, "pipeline")
        except Exception:
            pass

    def _merge(rr: int, first_col: int, last_col: int) -> None:
        if last_col <= first_col:
            return
        try:
            ws.merge_cells(start_row=rr, start_column=first_col, end_row=rr, end_column=last_col)
        except Exception:
            pass

    def _section(title: str, source_note: str = "") -> int:
        nonlocal row
        _merge(row, c1, c2)
        cell = ws.cell(row=row, column=c1, value=_safe_text(title, max_chars=140))
        cell.fill = section_fill
        cell.font = title_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        if source_note:
            _comment(cell, source_note)
        for cc in range(c1, c2 + 1):
            ws.cell(row=row, column=cc).fill = section_fill
            ws.cell(row=row, column=cc).border = thin_border
        _set_row_height_min(row, 19.5)
        out = row
        row += 1
        return out

    def _header(labels: Sequence[str], spans: Sequence[Tuple[int, int]]) -> int:
        nonlocal row
        for cc in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=cc, value="")
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        for idx, label in enumerate(labels):
            if idx >= len(spans):
                break
            s, e = spans[idx]
            _merge(row, s, e)
            ws.cell(row=row, column=s, value=_safe_text(label, max_chars=80))
        _set_row_height_min(row, 19.5)
        out = row
        row += 1
        return out

    def _row_values(values: Sequence[Any], spans: Sequence[Tuple[int, int]], *, fill: Optional[Any] = None, wrap_cols: Set[int] = frozenset(), source_note: str = "") -> int:
        nonlocal row
        row_fill = fill or neutral_fill
        for cc in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=cc)
            cell.fill = copy(row_fill)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        for idx, value in enumerate(values):
            if idx >= len(spans):
                break
            s, e = spans[idx]
            _merge(row, s, e)
            cleaned_value = _safe_text(value, max_chars=180) if isinstance(value, str) else value
            cell = ws.cell(row=row, column=s, value=cleaned_value)
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=s in wrap_cols,
            )
        if source_note:
            _comment(ws.cell(row=row, column=spans[-1][0] if spans else c1), source_note)
        _set_row_height_min(row, 19.5)
        out = row
        row += 1
        return out

    def _blank(height: float = 6.0, *, style_panel: bool = False) -> None:
        nonlocal row
        if style_panel:
            for cc in range(c1, c2 + 1):
                cell = ws.cell(row=row, column=cc)
                cell.fill = neutral_fill
                cell.border = spacer_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        _set_row_height_min(row, height)
        row += 1

    # Clear only the intentional side-panel range. The wider ANF scrubber should
    # already have removed stale O:AC blocks; this keeps the rewrite idempotent.
    # Keep this scrub local to the intentional top-right ANF side-panel.  The
    # lower Valuation section reuses O:V for output interpretation, market
    # pricing and convertible notes; clearing to max_row splits that PBI-style
    # layout and removes useful sections.
    _clear_panel_rect(1, 125, c1, c2)
    for col_letter, width in {
        get_column_letter(c1): 18,
        get_column_letter(c1 + 1): 16,
        get_column_letter(c1 + 2): 14,
        get_column_letter(c1 + 3): 16,
        get_column_letter(c1 + 4): 16,
        get_column_letter(c1 + 5): 16,
        get_column_letter(c1 + 6): 16,
        get_column_letter(c1 + 7): 16,
        get_column_letter(c1 + 8): 16,
        get_column_letter(c1 + 9): 16,
        get_column_letter(c1 + 10): 16,
        get_column_letter(c1 + 11): 16,
        get_column_letter(c1 + 12): 18,
        get_column_letter(c1 + 13): 18,
        get_column_letter(c1 + 14): 18,
    }.items():
        ws.column_dimensions[col_letter].width = width

    markers: Dict[str, int] = {}
    guidance_spans = [(c1, c1 + 1), (c1 + 2, c1 + 2), (c1 + 3, c1 + 3), (c1 + 4, c1 + 11), (c1 + 12, c2)]

    markers["guidance_latest"] = _section(
        "Guidance (As of 2026-01-31) - Status: Open",
        "2025-Q4 earnings release and normalized ANF guidance rows. Found: revenue, margin, EPS, buybacks, shares, capex, stores, tariffs.",
    )
    _header(["Metric", "Stated in", "Applies to", "Guidance", "Trend / realized"], guidance_spans)
    latest_rows = [
        ("Revenue growth", "2025-Q4", "2026 year", "+3-5%", "Open / not yet realized"),
        ("Operating margin", "2025-Q4", "2026 year", "12.0-12.5%", "Open; 2025: 13.3% / 12.5% adj"),
        ("Adjusted EPS", "2025-Q4", "2026 year", "$10.20-$11.00", "Open; 2025: $10.46 GAAP / $9.86 adj"),
        ("Buybacks", "2025-Q4", "2026 year", "around $450m", "Open; 2025 actual ~$450m"),
        ("Diluted shares", "2025-Q4", "2026 year", "around 45m", "Open / not yet realized"),
        ("Capex", "2025-Q4", "2026 year", "$200-$225m", "Open; 2025 actual $240.8m"),
        ("Store plan", "2025-Q4", "2026 year", "55 open / 25 close / 70 remodel", "Open / not yet realized"),
        ("Tariff headwind", "2025-Q4", "2026 year", "~70 bps / ~$40m headwind", "Open; bps bridge item"),
        ("Marketing headwind", "2025-Q4", "2026 year", "~50 bps higher as % sales", "Open; bps bridge item"),
        ("Q1 sales growth", "2025-Q4", "2026-Q1", "+1-3%", "Open / not yet realized"),
        ("Q1 operating margin", "2025-Q4", "2026-Q1", "around 7%", "Open / not yet realized"),
        ("Q1 Adjusted EPS", "2025-Q4", "2026-Q1", "$1.20-$1.30", "Open / not yet realized"),
        ("Q1 buybacks", "2025-Q4", "2026-Q1", "at least $100m", "Open / not yet realized"),
        ("Q1 diluted shares", "2025-Q4", "2026-Q1", "around 46m", "Open / not yet realized"),
        ("Q1 tariff headwind", "2025-Q4", "2026-Q1", "~290 bps / ~$30m headwind", "Open; bps bridge item"),
        ("Q1 freight tailwind", "2025-Q4", "2026-Q1", "~160 bps tailwind", "Open; bps bridge item"),
        ("Q1 ERP disruption", "2025-Q4", "2026-Q1", ">100 bps headwind", "Open; bps bridge item"),
    ]
    for idx, rec in enumerate(latest_rows):
        _row_values(rec, guidance_spans, fill=neutral_alt_fill if idx % 2 == 0 else neutral_fill, source_note="2025-Q4 earnings release / guidance table.")
    _blank()

    markers["guidance_prior"] = _section(
        "Guidance (As of 2025-11-01) - Status: Mixed",
        "2025-Q3 earnings release and normalized ANF guidance rows. Found: revenue, margin, EPS, buybacks, shares, capex, tariffs.",
    )
    _header(["Metric", "Stated in", "Applies to", "Guidance", "Trend / realized"], guidance_spans)
    prior_rows = [
        ("Revenue growth", "2025-Q3", "2025 year", "+6-7%", "Final 2025: +6%"),
        ("Operating margin", "2025-Q3", "2025 year", "13.0-13.5%", "Final 13.3% / 12.5% adj"),
        ("EPS / adjusted EPS", "2025-Q3", "2025 year", "$10.20-$10.50", "$10.46 GAAP; $9.86 adj"),
        ("Buybacks", "2025-Q3", "2025 year", "around $450m", "Actual around $450m"),
        ("Diluted shares", "2025-Q3", "2025 year", "around 48m", "Avg 48.5m; 2025-Q4 46.8m"),
        ("Capex", "2025-Q3", "2025 year", "around $225m", "Actual $240.8m"),
        ("Tariff impact", "2025-Q3", "2025 year", "around $90m cost pressure", "Bridge item; not rev guide"),
    ]
    for idx, rec in enumerate(prior_rows):
        _row_values(rec, guidance_spans, fill=neutral_alt_fill if idx % 2 == 0 else neutral_fill, source_note="2025-Q3 earnings release / guidance table.")
    _blank()

    markers["operating_drivers"] = _section("Operating Drivers", "ANF-specific driver map from Operating_Drivers and ANF source materials.")
    driver_spans = [(c1, c1 + 2), (c1 + 3, c1 + 5), (c1 + 6, c1 + 11), (c1 + 12, c2)]
    _header(["Driver group", "Driver", "Why it matters", "Source/type"], driver_spans)
    driver_rows = [
        ("Brand / demand", "A&F + Hollister", "Hollister growth engine and Abercrombie stabilization", "earnings release / quarterly history"),
        ("Comps / lapping", "brand/geography comps", "Separates true slowdown from tough compares", "quarterly history"),
        ("Margin bridge", "Tariff / freight / ERP", "Key debate for 2026 EBIT and EPS", "earnings release / transcript"),
        ("Inventory risk", "Inventory vs sales", "Retail risk indicator and margin risk", "earnings release / transcript"),
        ("Digital / omni", "44% digital / >1bn visits", "Supports omnichannel model quality", "annual report / transcript"),
        ("Stores", "open/close/remodel", "Growth and productivity driver", "quarterly history / guidance"),
        ("Capital allocation", "buybacks / net cash", "EPS support and downside buffer", "cash flow / guidance"),
        ("Geography", "Americas / EMEA / APAC", "International growth and regional risk", "segment data"),
    ]
    for idx, rec in enumerate(driver_rows):
        _row_values(rec, driver_spans, fill=neutral_alt_fill if idx % 2 == 0 else neutral_fill, wrap_cols={c1 + 3}, source_note="Operating_Drivers / source-backed ANF retail driver.")
    _blank()

    markers["thesis_bridge"] = _section("Thesis Bridge", "Quick ANF valuation bridge; detailed scenario work remains on ANF_Investment_Case.")
    thesis_spans = [(c1, c1 + 5), (c1 + 6, c1 + 8), (c1 + 9, c2)]
    thesis_value_col = thesis_spans[1][0]
    _row_values(["Quick valuation bridge; no market price required.", "", ""], thesis_spans, fill=neutral_alt_fill, source_note="Valuation input rows plus ANF guidance.")
    _header(["Bridge item", "Value", "Notes"], thesis_spans)
    thesis_inputs = [
        ("Adj EBITDA TTM", 815.590, "Adjusted metrics base $815.6m", "#,##0.0"),
        ("FCF TTM", 378.368, "CFO - capex $378.4m", "#,##0.0"),
        ("Net cash incl. securities", 784.576, "Cash + securities $784.6m", "#,##0.0"),
        ("Diluted shares", 46.837, "Latest diluted shares", "#,##0.0"),
        ("2026 EPS guide midpoint", 10.60, "$10.20-$11.00 midpoint", "$0.00"),
        ("P/E multiple", 13.0, "User input default", "0.0x"),
        ("EV/Adj EBITDA multiple", 8.0, "User input default", "0.0x"),
        ("FCF yield", 0.07, "User input default", "0.0%"),
        ("Lease-adjusted net debt", 408.555, "Supplemental lease view $408.6m", "#,##0.0"),
        ("Sales growth uplift / shortfall", 0.0, "User scenario input", "0.0%"),
        ("Margin uplift / shortfall", 0.0, "User scenario input", '0 "bps"'),
        ("Buyback/share adjustment", 0.0, "User scenario input, m shares", "#,##0.0"),
    ]
    thesis_row_by_label: Dict[str, int] = {}
    for idx, (label, value, note, num_fmt) in enumerate(thesis_inputs):
        rr = _row_values([label, value, note], thesis_spans, fill=neutral_alt_fill if idx % 2 == 0 else neutral_fill, source_note=note)
        thesis_row_by_label[label] = rr
        vcell = ws.cell(rr, thesis_value_col)
        vcell.number_format = num_fmt
        vcell.alignment = Alignment(horizontal="left", vertical="center")
        if label in {
            "P/E multiple",
            "EV/Adj EBITDA multiple",
            "FCF yield",
            "Sales growth uplift / shortfall",
            "Margin uplift / shortfall",
            "Buyback/share adjustment",
        }:
            vcell.fill = input_fill
            vcell.font = input_font
    _header(["Output", "Value", "Interpretation"], thesis_spans)
    base_ebitda_row = thesis_row_by_label["Adj EBITDA TTM"]
    fcf_row = thesis_row_by_label["FCF TTM"]
    net_cash_row = thesis_row_by_label["Net cash incl. securities"]
    shares_row = thesis_row_by_label["Diluted shares"]
    pe_row = thesis_row_by_label["P/E multiple"]
    ev_mult_row = thesis_row_by_label["EV/Adj EBITDA multiple"]
    fcf_yield_row = thesis_row_by_label["FCF yield"]
    eps_row = thesis_row_by_label["2026 EPS guide midpoint"]
    value_col_letter = get_column_letter(thesis_value_col)
    outputs = [
        ("Thesis Adj EBITDA", f"={value_col_letter}{base_ebitda_row}", "Adjusted EBITDA TTM base", "#,##0.0"),
        ("Thesis FCF", f"={value_col_letter}{fcf_row}", "FCF TTM base", "#,##0.0"),
        ("Thesis EPS", f"={value_col_letter}{eps_row}", "Guided midpoint", "$0.00"),
        ("EV @ EV/Adj EBITDA", f"={value_col_letter}{base_ebitda_row}*{value_col_letter}{ev_mult_row}", "Adjusted EBITDA × EV multiple", "#,##0.0"),
        ("Equity value @ EV/Adj EBITDA", f"={value_col_letter}{base_ebitda_row}*{value_col_letter}{ev_mult_row}+{value_col_letter}{net_cash_row}", "Uses core net cash", "#,##0.0"),
        ("Equity value @ P/E", f"={value_col_letter}{eps_row}*{value_col_letter}{pe_row}*{value_col_letter}{shares_row}", "EPS × P/E × diluted shares", "#,##0.0"),
        ("Equity value @ FCF yield", f"=IF({value_col_letter}{fcf_yield_row}<=0,\"\",{value_col_letter}{fcf_row}/{value_col_letter}{fcf_yield_row})", "FCF / yield", "#,##0.0"),
        ("__SPACER__", "", "", ""),
        ("Range summary", "", "Per-share output range", "$0"),
        ("Value/share @ P/E", f"={value_col_letter}{eps_row}*{value_col_letter}{pe_row}", "Guided EPS midpoint × P/E", "$0.00"),
        ("Value/share @ EV/Adj EBITDA", f"=({value_col_letter}{base_ebitda_row}*{value_col_letter}{ev_mult_row}+{value_col_letter}{net_cash_row})/{value_col_letter}{shares_row}", "Equity value divided by diluted shares", "$0.00"),
        ("Value/share @ FCF yield", f"=IF({value_col_letter}{fcf_yield_row}<=0,\"\",({value_col_letter}{fcf_row}/{value_col_letter}{fcf_yield_row})/{value_col_letter}{shares_row})", "FCF-yield value / shares", "$0.00"),
    ]
    thesis_audit_values: Dict[str, str] = {}
    try:
        base_adj_ebitda = 815.590
        base_fcf = 378.368
        base_net_cash = 784.576
        base_eps = 10.60
        base_shares = 46.837
        base_pe = 13.0
        base_ev_mult = 8.0
        base_fcf_yield = 0.07
        thesis_audit_values = {
            "Thesis EPS": f"${base_eps:.2f}",
            "Equity value @ P/E": f"${base_eps * base_pe * base_shares:,.1f}m",
            "Value/share @ P/E": f"${base_eps * base_pe:,.2f}",
            "EV @ EV/Adj EBITDA": f"${base_adj_ebitda * base_ev_mult:,.1f}m",
            "Equity value @ EV/Adj EBITDA": f"${base_adj_ebitda * base_ev_mult + base_net_cash:,.1f}m",
            "Value/share @ EV/Adj EBITDA": f"${(base_adj_ebitda * base_ev_mult + base_net_cash) / base_shares:,.2f}",
            "Equity value @ FCF yield": f"${base_fcf / base_fcf_yield:,.1f}m",
            "Value/share @ FCF yield": f"${(base_fcf / base_fcf_yield) / base_shares:,.2f}",
        }
    except Exception:
        thesis_audit_values = {}
    output_row_by_label: Dict[str, int] = {}
    visible_idx = 0
    for idx, (label, formula, note, num_fmt) in enumerate(outputs):
        if label == "__SPACER__":
            _blank(19.5, style_panel=True)
            continue
        rr = _row_values([label, formula, note], thesis_spans, fill=neutral_alt_fill if idx % 2 == 0 else neutral_fill, source_note=note)
        output_row_by_label[label] = rr
        vcell = ws.cell(rr, thesis_value_col)
        vcell.value = formula
        vcell.number_format = num_fmt
        vcell.alignment = Alignment(horizontal="left", vertical="center")
        if label in thesis_audit_values:
            _comment(vcell, f"Audit value before Excel recalculation: {thesis_audit_values[label]}. {note}")
        visible_idx += 1
    range_row = output_row_by_label.get("Range summary")
    value_share_rows = [
        output_row_by_label.get(label)
        for label in ("Value/share @ P/E", "Value/share @ EV/Adj EBITDA", "Value/share @ FCF yield")
        if output_row_by_label.get(label)
    ]
    if range_row and value_share_rows:
        refs = ",".join(f"{value_col_letter}{rr}" for rr in value_share_rows)
        ws.cell(range_row, thesis_value_col).value = (
            f"=TEXT(MIN({refs}),\"$0\")&\"-\"&TEXT(MAX({refs}),\"$0\")"
        )

    return markers
