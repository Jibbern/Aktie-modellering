"""Basis_Proxy_Sandbox writer helpers for the Excel economics overlay."""
from __future__ import annotations

import math
from copy import copy
from dataclasses import dataclass
from datetime import date
from typing import Any, Callable, Dict, List, Mapping, Optional, Sequence, Set, Tuple

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .excel_writer_market_data_sources import _convert_market_price_value
from .market_data.service import _gpre_official_market_weights_for_quarter


@dataclass(frozen=True)
class BasisProxySandboxWriterDeps:
    extract_operating_driver_rows_for_template: Callable[..., Any]
    load_operating_driver_source_records_by_quarter: Callable[..., Any]
    load_operating_driver_template_index: Callable[..., Any]
    market_quality_rank: Callable[..., Any]
    operating_driver_quarters: Callable[..., Any]
    overlay_coefficient_detail: Callable[..., Any]
    overlay_market_date_text: Callable[..., Any]
    quarter_label_short: Callable[..., Any]
    write_gpre_approx_market_crush_build_up_section: Callable[..., Any]
    align_center: Any
    align_center_wrap: Any
    analysis_theme: Any
    as_of_market_quarter: Any
    body_font: Any
    bold_font: Any
    border_color: Any
    coeff_rows: Any
    current_market_display_quarter: Any
    dark_text_color: Any
    economics_market_rows: Any
    font_size: Any
    gpre_commercial_setup_rows: Any
    gpre_plant_capacity_history: Any
    gpre_proxy_implied_results_bundle: Any
    gpre_reported_gallons_by_quarter: Any
    gpre_ticker_root_local: Any
    header_fill: Any
    header_size: Any
    intro_fill: Any
    is_gpre_profile: Any
    market_input_templates_by_key: Any
    market_rows: Any
    muted_text_color: Any
    next_thesis_quarter_end: Any
    prior_market_display_quarter: Any
    quarter_open_display_quarter: Any
    row_map: Any
    section_fill: Any
    thin_border: Any
    zebra_fill_dark: Any
    zebra_fill_light: Any

def write_basis_proxy_sandbox_sheet(
    deps: BasisProxySandboxWriterDeps,
    target_ws: Any,
    start_row: int,
    model_result: Dict[str, Any],
) -> Dict[str, Any]:
    _extract_operating_driver_rows_for_template = deps.extract_operating_driver_rows_for_template
    _load_operating_driver_source_records_by_quarter = deps.load_operating_driver_source_records_by_quarter
    _load_operating_driver_template_index = deps.load_operating_driver_template_index
    _market_quality_rank = deps.market_quality_rank
    _operating_driver_quarters = deps.operating_driver_quarters
    _overlay_coefficient_detail = deps.overlay_coefficient_detail
    _overlay_market_date_text = deps.overlay_market_date_text
    _quarter_label_short = deps.quarter_label_short
    _write_gpre_approx_market_crush_build_up_section = deps.write_gpre_approx_market_crush_build_up_section
    align_center = deps.align_center
    align_center_wrap = deps.align_center_wrap
    analysis_theme = deps.analysis_theme
    as_of_market_quarter = deps.as_of_market_quarter
    body_font = deps.body_font
    bold_font = deps.bold_font
    border_color = deps.border_color
    coeff_rows = deps.coeff_rows
    current_market_display_quarter = deps.current_market_display_quarter
    dark_text_color = deps.dark_text_color
    economics_market_rows = deps.economics_market_rows
    font_size = deps.font_size
    gpre_commercial_setup_rows = deps.gpre_commercial_setup_rows
    gpre_plant_capacity_history = deps.gpre_plant_capacity_history
    gpre_proxy_implied_results_bundle = deps.gpre_proxy_implied_results_bundle
    gpre_reported_gallons_by_quarter = deps.gpre_reported_gallons_by_quarter
    gpre_ticker_root_local = deps.gpre_ticker_root_local
    header_fill = deps.header_fill
    header_size = deps.header_size
    intro_fill = deps.intro_fill
    is_gpre_profile = deps.is_gpre_profile
    market_input_templates_by_key = deps.market_input_templates_by_key
    market_rows = deps.market_rows
    muted_text_color = deps.muted_text_color
    next_thesis_quarter_end = deps.next_thesis_quarter_end
    prior_market_display_quarter = deps.prior_market_display_quarter
    quarter_open_display_quarter = deps.quarter_open_display_quarter
    row_map = deps.row_map
    section_fill = deps.section_fill
    thin_border = deps.thin_border
    zebra_fill_dark = deps.zebra_fill_dark
    zebra_fill_light = deps.zebra_fill_light

    if not isinstance(model_result, dict):
        return {}
    quarterly_df = model_result.get("quarterly_df")
    metrics_df = model_result.get("metrics_df")
    leaderboard_df = model_result.get("leaderboard_df")
    weights_df = model_result.get("weights_df")
    footprint_df = model_result.get("footprint_df")
    hedge_style_study = model_result.get("hedge_style_study") if isinstance(model_result.get("hedge_style_study"), dict) else {}
    bid_adjusted_offsets_df = model_result.get("bid_adjusted_offsets_df")
    gpre_bid_snapshot = model_result.get("gpre_bid_snapshot")
    recommended_model_key = str(model_result.get("recommended_model_key") or "")
    incumbent_baseline_model_key = str(model_result.get("incumbent_baseline_model_key") or "")
    expanded_best_candidate_model_key = str(model_result.get("expanded_best_candidate_model_key") or model_result.get("expanded_candidate_model_key") or "")
    best_historical_fit_model_key = str(model_result.get("best_historical_fit_model_key") or "")
    best_compromise_model_key = str(model_result.get("best_compromise_model_key") or "")
    best_forward_lens_model_key = str(model_result.get("best_forward_lens_model_key") or "")
    production_winner_model_key = str(model_result.get("production_winner_model_key") or model_result.get("gpre_proxy_model_key") or "")
    promotion_guard_reason = str(model_result.get("gpre_proxy_promotion_guard_reason") or "")
    production_decision_story = str(model_result.get("production_decision_story") or "").strip()
    selection_vs_promotion_explanation = str(model_result.get("selection_vs_promotion_explanation") or "").strip()
    summary_md = str(model_result.get("summary_markdown") or "").strip()
    experimental_signal_audit = model_result.get("experimental_signal_audit") if isinstance(model_result.get("experimental_signal_audit"), dict) else {}
    experimental_candidate_comparison_df = (
        model_result.get("experimental_candidate_comparison_df")
        if isinstance(model_result.get("experimental_candidate_comparison_df"), pd.DataFrame)
        else pd.DataFrame()
    )
    coproduct_experimental_candidate_comparison_df = (
        model_result.get("coproduct_experimental_candidate_comparison_df")
        if isinstance(model_result.get("coproduct_experimental_candidate_comparison_df"), pd.DataFrame)
        else pd.DataFrame()
    )
    coproduct_experimental_method_specs = [
        dict(rec)
        for rec in list(model_result.get("coproduct_experimental_method_specs") or [])
        if isinstance(rec, dict)
    ]
    coproduct_experimental_legacy_reference_model_key = str(
        model_result.get("coproduct_experimental_legacy_reference_model_key") or ""
    )
    coproduct_experimental_legacy_reference_row = (
        dict(model_result.get("coproduct_experimental_legacy_reference_row") or {})
        if isinstance(model_result.get("coproduct_experimental_legacy_reference_row"), dict)
        else {}
    )
    best_coproduct_experimental_historical_model_key = str(
        model_result.get("best_coproduct_experimental_historical_model_key") or ""
    )
    best_coproduct_experimental_compromise_model_key = str(
        model_result.get("best_coproduct_experimental_compromise_model_key") or ""
    )
    best_coproduct_experimental_forward_model_key = str(
        model_result.get("best_coproduct_experimental_forward_model_key") or ""
    )
    best_coproduct_experimental_model_key = str(
        model_result.get("best_coproduct_experimental_model_key")
        or best_coproduct_experimental_compromise_model_key
        or ""
    )
    coproduct_experimental_frame_values = (
        dict(model_result.get("coproduct_experimental_frame_values") or {})
        if isinstance(model_result.get("coproduct_experimental_frame_values"), dict)
        else {}
    )
    coproduct_experimental_summary_md = str(
        model_result.get("coproduct_experimental_summary_markdown") or ""
    ).strip()
    ws = target_ws
    overlay_source_ws = (
        target_ws.parent["Economics_Overlay"]
        if "Economics_Overlay" in list(getattr(target_ws.parent, "sheetnames", []) or [])
        else None
    )
    approx_market_crush_build_up_layout: Dict[str, Any] = {}
    corn_oil_gate_check_layout: Dict[str, Any] = {}
    coproduct_frame_summary_layout: Dict[str, Any] = {}
    coproduct_signal_readiness_layout: Dict[str, Any] = {}
    coproduct_volume_support_layout: Dict[str, Any] = {}
    coproduct_experimental_layout: Dict[str, Any] = {}

    def _sandbox_model_label(model_key_in: Any) -> str:
        key_txt = str(model_key_in or "").strip()
        return {
            "simple_market": "Simple market",
            "bridge_current_quarter_avg": "Bridge current avg",
            "bridge_front_loaded": "Bridge front-loaded",
            "bridge_current75_prev25": "Bridge 75/25",
            "bridge_current50_prev50": "Bridge 50/50",
            "process_current_quarter_avg": "Process current avg",
            "process_front_loaded": "Process front-loaded",
            "process_current75_prev25": "Process 75/25",
            "process_current50_prev50": "Process 50/50",
            "process_quarter_open_blend": "Process q-open blend",
            "process_quarter_open_blend_ops_penalty": "Process q-open blend + ops penalty",
            "process_quarter_open_blend_hedge_realization": "Process q-open + hedge realization",
            "process_quarter_open_blend_exec_penalty": "Process q-open + severe ops penalty",
            "process_quarter_open_blend_utilization_penalty": "Process q-open + utilization penalty",
            "process_quarter_open_blend_maintenance_delay_penalty": "Process q-open + maintenance delay",
            "process_quarter_open_blend_inventory_timing_drag": "Process q-open + inventory drag",
            "process_quarter_open_blend_locked_setup": "Process q-open + locked setup",
            "process_basis_blend_current40_front60": "Process basis blend 40/60",
            "process_basis_passthrough_beta35": "Process basis beta 0.35",
            "process_basis_passthrough_beta65": "Process basis beta 0.65",
            "process_quarter_open_current50_exec_penalty": "Process q-open/current 50/50 + exec penalty",
            "process_regime_basis_passthrough": "Process regime basis passthrough",
            "process_two_stage_realization_residual": "Process two-stage residual",
            "process_capacity_weighted_basis_strict": "Process capacity-weighted basis strict",
            "process_inventory_gap_penalty_small": "Process inventory gap penalty small",
            "process_inventory_gap_penalty_medium": "Process inventory gap penalty medium",
            "process_utilization_regime_blend": "Process utilization regime blend",
            "process_utilization_regime_residual": "Process utilization regime residual",
            "process_exec_inventory_combo_medium": "Process exec + inventory combo",
            "process_asymmetric_basis_passthrough": "Process asymmetric basis passthrough",
            "process_market_process_ensemble_35_65": "Market/process ensemble 35/65",
            "process_locked_share_asymmetric_passthrough": "Locked-share asymmetric passthrough",
            "process_prior_gap_carryover_small": "Prior-gap carryover small",
            "process_prior_disturbance_carryover": "Prior-disturbance carryover",
            "process_residual_regime_locked_vs_disturbed": "Process residual regime split",
            "process_gated_incumbent_vs_residual": "Process gated incumbent vs residual",
            "simple_plus_10pct_credit": "Simple + 10% credit",
            "simple_plus_15pct_credit": "Simple + 15% credit",
            "simple_plus_20pct_credit": "Simple + 20% credit",
            "simple_plus_25pct_credit": "Simple + 25% credit",
            "simple_plus_30pct_credit": "Simple + 30% credit",
            "simple_plus_10pct_coverage_credit": "Simple + 10% coverage credit",
            "simple_plus_20pct_coverage_credit": "Simple + 20% coverage credit",
            "simple_plus_30pct_coverage_credit": "Simple + 30% coverage credit",
            "simple_plus_25pct_credit_less_2c": "Simple + 25% credit less 2c",
            "simple_plus_30pct_coverage_credit_less_2c": "Simple + 30% coverage credit less 2c",
            "simple_plus_half_credit": "Simple + 50% credit",
            "process_front_loaded_ops_penalty": "Process front + ops penalty",
            "process_front_loaded_ethanol_geo": "Process front + ethanol geo",
            "spot_simple": "Spot simple",
            "quarter_open_lock_25": "Quarter-open lock 25%",
            "quarter_open_lock_50": "Quarter-open lock 50%",
            "quarter_open_lock_75": "Quarter-open lock 75%",
            "front_loaded_layering": "Front-loaded layering",
            "equal_monthly_layering": "Equal monthly layering",
            "quarter_open_plus_current_blend": "Quarter-open + current blend",
            "good_setup_realization_drag": "Good setup + realization drag",
            "ops_disruption_overlay": "Ops disruption overlay",
            "hedge_disclosed_bridge_prior_current": "Disclosed memo bridge prior-current",
            "hedge_disclosed_bridge_prior_front": "Disclosed memo bridge prior-front",
            "hedge_disclosed_process_prior_current": "Disclosed memo process prior-current",
            "hedge_disclosed_process_prior_front": "Disclosed memo process prior-front",
            "hedge_pattern_bridge_prior_current": "Pattern memo bridge prior-current",
            "hedge_pattern_bridge_prior_front": "Pattern memo bridge prior-front",
            "hedge_pattern_process_prior_current": "Pattern memo process prior-current",
            "hedge_pattern_process_prior_front": "Pattern memo process prior-front",
            "bid_adjusted_offset": "Bid-adjusted offset",
            "plant_count_weighted": "Plant-count weighted",
            "plant_count_front_loaded": "Plant-count front-loaded",
            "plant_count_prev_quarter": "Plant-count prior-quarter lag",
            "plant_count_bid_adjusted_offset": "Plant-count bid-adjusted offset",
            "capacity_weighted": "Capacity-weighted",
            "equal_weighted": "Equal-weight",
            "optimized_weights": "Optimized weights",
            "baseline_market_proxy": "Baseline market proxy",
            "calibrated_sensitivity": "Calibrated sensitivity",
        }.get(key_txt, key_txt.replace("_", " ").title())

    def _sandbox_split_label(split_in: Any) -> str:
        split_txt = str(split_in or "").strip()
        return {
            "train": "Train",
            "test": "Test",
            "full": "Full",
            "clean_reported_window": "Clean reported",
            "diag_underlying": "Underlying diag",
        }.get(split_txt, split_txt.replace("_", " ").title())

    def _sandbox_short_regime_flags(text_in: Any) -> str:
        pieces = [str(part or "").strip() for part in str(text_in or "").split(";") if str(part or "").strip()]
        mapping = {
            "fairmont_active": "Fairmont active",
            "fairmont_offline": "Fairmont offline",
            "obion_active": "Obion active",
            "partial_tennessee_coverage": "Tennessee partial coverage",
            "post_obion_sale": "Post-Obion sale",
            "eight_plant_footprint": "8-plant footprint",
            "pre_fairmont_idling": "Pre-Fairmont idling",
        }
        display_pieces = [mapping.get(piece, piece.replace("_", " ")) for piece in pieces]
        return "; ".join(display_pieces)

    def _sandbox_short_coverage(text_in: Any) -> str:
        low = str(text_in or "").strip().lower()
        if not low:
            return ""
        if "unsupported basis coverage excluded" in low and "tennessee" in low:
            return "Tennessee excluded from primary basis coverage"
        return str(text_in or "").strip()

    def _sandbox_short_denominator(text_in: Any) -> str:
        low = str(text_in or "").strip().lower()
        return (
            "Gallons sold" if low == "ethanol gallons sold"
            else "Gallons produced" if low == "ethanol gallons produced"
            else "Estimated gallons (corn x yield)" if "estimated gallons" in low
            else str(text_in or "").strip()
        )

    def _sandbox_active_regions(text_in: Any) -> str:
        pieces = [str(part or "").strip() for part in str(text_in or "").split(",") if str(part or "").strip()]
        return ", ".join(piece.replace("_", " ").title() for piece in pieces)

    def _sandbox_guard_label(reason_in: Any) -> str:
        reason_txt = str(reason_in or "").strip()
        return {
            "passed_guardrails": "Passed selection guardrails",
            "clean_mae_exceeded_best_window_tolerance": "Selection blocked: clean tolerance",
            "q1_mae_exceeded_best_window_tolerance": "Selection blocked: Q1 tolerance",
            "q1_mean_error_exceeded_bias_limit": "Selection blocked: Q1 bias",
            "incumbent_baseline": "Incumbent baseline",
            "passed_promotion_guardrails": "Passed promotion guardrails",
            "preview_support_incomplete": "Promotion blocked: preview incomplete",
            "live_preview_quality_not_faithful_enough": "Promotion blocked: preview not faithful enough",
            "incremental_distance_vs_official_too_low": "Promotion blocked: low distance vs official",
            "too_few_material_diff_quarters_vs_official": "Promotion blocked: too few diff quarters",
            "clean_mae_exceeded_incumbent_tolerance": "Promotion blocked: clean tolerance",
            "q1_mae_exceeded_incumbent_tolerance": "Promotion blocked: Q1 tolerance",
            "mean_error_exceeded_tolerance": "Promotion blocked: mean-error tolerance",
            "incumbent_improvement_threshold_not_met": "Promotion blocked: improvement threshold",
            "hard_quarter_mae_materially_worse_than_incumbent": "Promotion blocked: hard-quarter regression",
            "not_new_candidate": "Promotion blocked: not a new candidate",
            "expanded_best_is_incumbent_baseline": "Expanded best already incumbent",
            "promoted_over_incumbent_baseline": "Promoted over incumbent",
            "incumbent_baseline_missing": "Incumbent baseline missing",
            "promoted_no_incumbent_baseline_available": "Promoted with no incumbent baseline",
        }.get(reason_txt, reason_txt.replace("_", " "))

    def _sandbox_preview_phase_label(phase_in: Any) -> str:
        phase_txt = str(phase_in or "").strip()
        return {
            "prior": "prior",
            "quarter_open": "quarter-open",
            "current": "current",
            "next": "next",
        }.get(phase_txt, phase_txt.replace("_", "-"))

    overlay_prior_col = 2
    overlay_quarter_open_col = 4 if (is_gpre_profile and gpre_commercial_setup_rows) else 2
    overlay_current_col = 6 if (is_gpre_profile and gpre_commercial_setup_rows) else 2
    overlay_next_col = 8 if (is_gpre_profile and gpre_commercial_setup_rows) else 2

    def _empty_coproduct_state() -> Dict[str, bool]:
        return {
            "historical": False,
            "current": False,
            "next": False,
        }

    def _overlay_cell_has_explicit_numeric_value(row_in: Any, col_in: Any) -> bool:
        if overlay_source_ws is None or not isinstance(row_in, int) or not isinstance(col_in, int):
            return False
        if row_in <= 0 or col_in <= 0:
            return False
        raw_val = overlay_source_ws.cell(row=row_in, column=col_in).value
        if raw_val in (None, ""):
            return False
        if isinstance(raw_val, str):
            txt = str(raw_val or "").strip()
            if not txt or txt.startswith("="):
                return False
        num_val = pd.to_numeric(raw_val, errors="coerce")
        return pd.notna(num_val)

    def _overlay_row_readiness_state(row_in: Any) -> Dict[str, bool]:
        if not isinstance(row_in, int) or row_in <= 0:
            return _empty_coproduct_state()
        return {
            "historical": _overlay_cell_has_explicit_numeric_value(row_in, overlay_prior_col),
            "current": (
                _overlay_cell_has_explicit_numeric_value(row_in, overlay_quarter_open_col)
                or _overlay_cell_has_explicit_numeric_value(row_in, overlay_current_col)
            ),
            "next": _overlay_cell_has_explicit_numeric_value(row_in, overlay_next_col),
        }

    def _overlay_constant_input_state(row_in: Any) -> Dict[str, bool]:
        ready = _overlay_cell_has_explicit_numeric_value(row_in, 2)
        return {
            "historical": ready,
            "current": ready,
            "next": ready,
        }

    def _combine_coproduct_state(*states_in: Dict[str, bool]) -> Dict[str, bool]:
        out = _empty_coproduct_state()
        for bucket in out:
            out[bucket] = any(bool((state or {}).get(bucket)) for state in states_in)
        return out

    def _require_coproduct_state(*states_in: Dict[str, bool]) -> Dict[str, bool]:
        out = _empty_coproduct_state()
        for bucket in out:
            out[bucket] = bool(states_in) and all(bool((state or {}).get(bucket)) for state in states_in)
        return out

    def _coproduct_source_state(*, source_type_prefix: str, series_prefixes: Tuple[str, ...]) -> Dict[str, bool]:
        out = _empty_coproduct_state()
        quarter_anchor = as_of_market_quarter if isinstance(as_of_market_quarter, date) else None
        for rec in economics_market_rows:
            source_type_txt = str(rec.get("source_type") or "").strip().lower()
            if not source_type_txt.startswith(str(source_type_prefix or "").strip().lower()):
                continue
            series_key_txt = str(rec.get("series_key") or "").strip().lower()
            if not any(series_key_txt.startswith(prefix) for prefix in series_prefixes):
                continue
            rec_q = rec.get("quarter")
            if not isinstance(rec_q, date):
                continue
            if quarter_anchor is None:
                out["historical"] = True
                continue
            if rec_q < quarter_anchor:
                out["historical"] = True
            elif rec_q == quarter_anchor:
                out["current"] = True
            elif rec_q > quarter_anchor:
                out["next"] = True
        return out

    def _overlay_source_mode_text(row_in: Any) -> str:
        if overlay_source_ws is None or not isinstance(row_in, int) or row_in <= 0:
            return ""
        return str(overlay_source_ws.cell(row=row_in, column=11).value or "").strip()

    def _classify_coproduct_resolved_source(*source_texts_in: Any) -> str:
        labels: List[str] = []
        for source_txt in source_texts_in:
            low = str(source_txt or "").strip().lower()
            if not low:
                continue
            if "ams_3618" in low:
                label = "AMS 3618"
            elif "nwer" in low:
                label = "NWER"
            else:
                continue
            if label not in labels:
                labels.append(label)
        if not labels:
            return "Unknown/blank"
        if len(labels) == 1:
            return labels[0]
        return "Mixed"

    def _gpre_proxy_implied_frame_record(frame_key: str) -> Dict[str, Any]:
        if not (is_gpre_profile and gpre_commercial_setup_rows):
            return {}
        frame_map = (gpre_proxy_implied_results_bundle or {}).get("frames") or {}
        frame = frame_map.get(str(frame_key or "")) if isinstance(frame_map, dict) else {}
        return dict(frame) if isinstance(frame, dict) else {}

    def _parse_quarter_label_text(value_in: Any) -> Optional[date]:
        if isinstance(value_in, date):
            return value_in
        txt = str(value_in or "").strip()
        match = re.fullmatch(r"(\d{4})-Q([1-4])", txt)
        if not match:
            return None
        year_num = int(match.group(1))
        quarter_num = int(match.group(2))
        quarter_end_map = {
            1: date(year_num, 3, 31),
            2: date(year_num, 6, 30),
            3: date(year_num, 9, 30),
            4: date(year_num, 12, 31),
        }
        return quarter_end_map.get(quarter_num)

    def _safe_int_like(value_in: Any, default: int = 0) -> int:
        numeric = pd.to_numeric(value_in, errors="coerce")
        if pd.isna(numeric):
            return int(default)
        return int(numeric)

    coproduct_region_series_candidates: Dict[str, Dict[str, Tuple[str, ...]]] = {
        "renewable_corn_oil_price": {
            "nebraska": ("corn_oil_nebraska",),
            "illinois": ("corn_oil_illinois",),
            "indiana": ("corn_oil_indiana",),
            "iowa_east": ("corn_oil_iowa_east", "corn_oil_iowa_avg", "corn_oil_eastern_cornbelt"),
            "iowa_west": ("corn_oil_iowa_west", "corn_oil_iowa_avg"),
            "minnesota": ("corn_oil_minnesota",),
            "tennessee": tuple(),
        },
        "distillers_grains_price": {
            "nebraska": ("ddgs_10_nebraska",),
            "illinois": ("ddgs_10_illinois",),
            "indiana": ("ddgs_10_indiana",),
            "iowa_east": ("ddgs_10_iowa_east", "ddgs_10_iowa"),
            "iowa_west": ("ddgs_10_iowa_west", "ddgs_10_iowa"),
            "minnesota": ("ddgs_10_minnesota",),
            "tennessee": tuple(),
        },
    }
    coproduct_source_priority = ("nwer", "ams_3618")
    distillers_yield_num = pd.to_numeric((_overlay_coefficient_detail("distillers_yield") or {}).get("value"), errors="coerce")
    uhp_yield_num = pd.to_numeric((_overlay_coefficient_detail("uhp_yield") or {}).get("value"), errors="coerce")
    corn_oil_yield_num = pd.to_numeric((_overlay_coefficient_detail("renewable_corn_oil_yield") or {}).get("value"), errors="coerce")
    ethanol_yield_num = pd.to_numeric((_overlay_coefficient_detail("ethanol_yield") or {}).get("value"), errors="coerce")
    historical_gallons_million_map: Dict[date, float] = {}
    for raw_qd, raw_val in dict(gpre_reported_gallons_by_quarter or {}).items():
        qd = _parse_quarter_label_text(raw_qd)
        gallons_num = pd.to_numeric(raw_val, errors="coerce")
        if qd is None or pd.isna(gallons_num):
            continue
        gallons_float = float(gallons_num)
        if not math.isfinite(gallons_float) or gallons_float <= 0.0:
            continue
        historical_gallons_million_map[qd] = gallons_float
    coproduct_volume_support_specs: Dict[str, Dict[str, str]] = {
        "distillers_grains": {
            "label": "Distillers grains volume",
            "source_path": "Operating_Drivers exact filing-text parse (k tons)",
            "historical_usable": "YES",
            "current_usable": "YES - latest actual intensity anchor",
            "next_usable": "YES - latest actual intensity anchor",
            "best_use": "Forward coproduct yield / $m anchor",
        },
        "renewable_corn_oil": {
            "label": "Renewable corn oil volume",
            "source_path": "Operating_Drivers exact filing-text parse (million lbs)",
            "historical_usable": "YES",
            "current_usable": "YES - latest actual intensity anchor",
            "next_usable": "YES - latest actual intensity anchor",
            "best_use": "Forward coproduct yield / $m anchor",
        },
        "ultra_high_protein": {
            "label": "Ultra-high protein volume",
            "source_path": "Operating_Drivers exact filing-text parse (k tons)",
            "historical_usable": "YES",
            "current_usable": "YES - latest actual intensity anchor",
            "next_usable": "YES - latest actual intensity anchor",
            "best_use": "Secondary forward mix / yield anchor",
        },
        "protein_coproduct_mix": {
            "label": "Protein / coproduct mix commentary",
            "source_path": "Operating_Drivers commentary extract",
            "historical_usable": "Commentary only",
            "current_usable": "NO",
            "next_usable": "NO",
            "best_use": "Context only",
        },
    }
    operating_driver_volume_rows_cache: Dict[str, Dict[date, Dict[str, Any]]] = {}
    operating_driver_volume_ratio_cache: Dict[str, Dict[str, Any]] = {}
    coproduct_latest_yield_anchor_cache: Dict[str, Dict[str, Any]] = {}
    coproduct_volume_support_records_cache: Optional[List[Dict[str, Any]]] = None

    def _operating_driver_volume_rows(driver_key: str) -> Dict[date, Dict[str, Any]]:
        cache_key = str(driver_key or "").strip().lower()
        cached = operating_driver_volume_rows_cache.get(cache_key)
        if isinstance(cached, dict):
            return {qd: dict(rec) for qd, rec in cached.items()}
        template_by_key = dict((_load_operating_driver_template_index() or {}).get("template_by_key") or {})
        tpl = template_by_key.get(cache_key)
        if tpl is None:
            operating_driver_volume_rows_cache[cache_key] = {}
            return {}
        source_records_by_quarter = _load_operating_driver_source_records_by_quarter()
        out: Dict[date, Dict[str, Any]] = {}
        for qd in _operating_driver_quarters():
            quarter_records = source_records_by_quarter.get(qd, [])
            extracted_rows = list(_extract_operating_driver_rows_for_template(qd, tpl, quarter_records=quarter_records) or [])
            if not extracted_rows:
                continue
            chosen = next(
                (
                    dict(row)
                    for row in extracted_rows
                    if pd.notna(pd.to_numeric(row.get("Value"), errors="coerce"))
                ),
                dict(extracted_rows[0]),
            )
            out[qd] = chosen
        operating_driver_volume_rows_cache[cache_key] = {qd: dict(rec) for qd, rec in out.items()}
        return out

    def _operating_driver_volume_ratio_summary(driver_key: str) -> Dict[str, Any]:
        cache_key = str(driver_key or "").strip().lower()
        cached = operating_driver_volume_ratio_cache.get(cache_key)
        if isinstance(cached, dict):
            return dict(cached)
        history_rows = _operating_driver_volume_rows(cache_key)
        ratio_map: Dict[date, float] = {}
        ratio_unit = ""
        for qd, rec in history_rows.items():
            value_num = pd.to_numeric(rec.get("Value"), errors="coerce")
            gallons_num = pd.to_numeric(historical_gallons_million_map.get(qd), errors="coerce")
            if pd.isna(value_num) or pd.isna(gallons_num) or float(gallons_num) <= 0.0:
                continue
            if cache_key in {"distillers_grains", "ultra_high_protein"}:
                ratio_map[qd] = float(value_num) * 1000.0 / float(gallons_num)
                ratio_unit = "tons/mm gal"
            elif cache_key == "renewable_corn_oil":
                ratio_map[qd] = float(value_num) / float(gallons_num)
                ratio_unit = "lbs/gal"
        ordered_quarters = sorted(ratio_map)
        trailing_quarters = ordered_quarters[-8:] if len(ordered_quarters) > 8 else ordered_quarters
        trailing_values = [float(ratio_map[qd]) for qd in trailing_quarters]
        latest_quarter = trailing_quarters[-1] if trailing_quarters else None
        latest_ratio = ratio_map.get(latest_quarter) if isinstance(latest_quarter, date) else None
        out = {
            "ratio_unit": ratio_unit,
            "latest_quarter": latest_quarter,
            "latest_quarter_label": (_quarter_label_short(latest_quarter) if isinstance(latest_quarter, date) else ""),
            "latest_ratio": latest_ratio,
            "recent_min": (min(trailing_values) if trailing_values else None),
            "recent_max": (max(trailing_values) if trailing_values else None),
            "recent_count": len(trailing_values),
        }
        operating_driver_volume_ratio_cache[cache_key] = dict(out)
        return out

    def _latest_coproduct_yield_anchor(target_quarter_end: Optional[date]) -> Dict[str, Any]:
        cache_key = target_quarter_end.isoformat() if isinstance(target_quarter_end, date) else "latest"
        cached = coproduct_latest_yield_anchor_cache.get(cache_key)
        if isinstance(cached, dict):
            return dict(cached)
        distillers_rows = _operating_driver_volume_rows("distillers_grains")
        uhp_rows = _operating_driver_volume_rows("ultra_high_protein")
        oil_rows = _operating_driver_volume_rows("renewable_corn_oil")
        candidate_quarters = sorted(
            {
                qd
                for qd in set(distillers_rows) | set(uhp_rows) | set(oil_rows) | set(historical_gallons_million_map)
                if isinstance(qd, date)
                and (not isinstance(target_quarter_end, date) or qd <= target_quarter_end)
            },
            reverse=True,
        )
        for qd in candidate_quarters:
            corn_num = pd.to_numeric((row_map.get(("corn_consumed", qd)) or {}).get("Value"), errors="coerce")
            if pd.isna(corn_num) or float(corn_num) <= 0.0:
                gallons_num = pd.to_numeric(historical_gallons_million_map.get(qd), errors="coerce")
                if pd.notna(gallons_num) and pd.notna(ethanol_yield_num) and float(ethanol_yield_num) > 0.0:
                    corn_num = float(gallons_num) / float(ethanol_yield_num)
            if pd.isna(corn_num) or float(corn_num) <= 0.0:
                continue
            distillers_num = pd.to_numeric((distillers_rows.get(qd) or {}).get("Value"), errors="coerce")
            uhp_num = pd.to_numeric((uhp_rows.get(qd) or {}).get("Value"), errors="coerce")
            oil_num = pd.to_numeric((oil_rows.get(qd) or {}).get("Value"), errors="coerce")
            anchor: Dict[str, Any] = {
                "anchor_quarter": qd,
                "anchor_quarter_label": _quarter_label_short(qd),
                "corn_consumed_mbu": float(corn_num),
            }
            if pd.notna(distillers_num) and float(distillers_num) > 0.0:
                anchor["distillers_yield_lbs_per_bu"] = float(distillers_num) * 2.0 / float(corn_num)
            if pd.notna(uhp_num) and float(uhp_num) > 0.0:
                anchor["uhp_yield_lbs_per_bu"] = float(uhp_num) * 2.0 / float(corn_num)
            if pd.notna(oil_num) and float(oil_num) > 0.0:
                anchor["renewable_corn_oil_yield_lbs_per_bu"] = float(oil_num) / float(corn_num)
            if any(key.endswith("_yield_lbs_per_bu") for key in anchor):
                anchor["source_mode"] = "latest_reported_coproduct_volume_intensity"
                anchor["note"] = f"Volume yields anchored to latest actual {_quarter_label_short(qd)} coproduct/corn volumes."
                coproduct_latest_yield_anchor_cache[cache_key] = dict(anchor)
                return anchor
        coproduct_latest_yield_anchor_cache[cache_key] = {}
        return {}

    def _coproduct_volume_audit_note(driver_key: str) -> str:
        key_txt = str(driver_key or "").strip().lower()
        if key_txt == "protein_coproduct_mix":
            return "Commentary only; distillers/UHP mentions provide mix context, not a numeric volume curve."
        ratio_summary = _operating_driver_volume_ratio_summary(key_txt)
        ratio_unit = str(ratio_summary.get("ratio_unit") or "").strip()
        recent_count = int(ratio_summary.get("recent_count") or 0)
        latest_label = str(ratio_summary.get("latest_quarter_label") or "").strip()
        latest_ratio = pd.to_numeric(ratio_summary.get("latest_ratio"), errors="coerce")
        recent_min = pd.to_numeric(ratio_summary.get("recent_min"), errors="coerce")
        recent_max = pd.to_numeric(ratio_summary.get("recent_max"), errors="coerce")
        if recent_count >= 4 and ratio_unit and pd.notna(latest_ratio) and pd.notna(recent_min) and pd.notna(recent_max):
            prefix = "Secondary forward mix anchor" if key_txt == "ultra_high_protein" else "Forward volume-intensity anchor"
            return (
                f"Latest 12q actuals only. {prefix} {latest_label} {float(latest_ratio):.3f} {ratio_unit} "
                f"vs recent 8q range {float(recent_min):.3f}-{float(recent_max):.3f}."
            )
        if key_txt == "ultra_high_protein":
            return "Latest 12q actuals only. Secondary mix context; use only when a sourced latest volume/corn anchor exists."
        return "Latest 12q actuals only. Use as a forward volume-intensity anchor only when a sourced latest volume/corn anchor exists."

    def _coproduct_volume_support_audit_records() -> List[Dict[str, Any]]:
        nonlocal coproduct_volume_support_records_cache
        if isinstance(coproduct_volume_support_records_cache, list):
            return [dict(rec) for rec in coproduct_volume_support_records_cache]
        records: List[Dict[str, Any]] = []
        for driver_key, spec in coproduct_volume_support_specs.items():
            records.append(
                {
                    "driver_key": driver_key,
                    "label": str(spec.get("label") or ""),
                    "source_path": str(spec.get("source_path") or ""),
                    "historical_usable": str(spec.get("historical_usable") or ""),
                    "current_usable": str(spec.get("current_usable") or ""),
                    "next_usable": str(spec.get("next_usable") or ""),
                    "best_use": str(spec.get("best_use") or ""),
                    "note": _coproduct_volume_audit_note(driver_key),
                }
            )
        coproduct_volume_support_records_cache = [dict(rec) for rec in records]
        return records
    weighted_coproduct_input_cache: Dict[Tuple[str, date, str], Dict[str, Any]] = {}
    weighted_coproduct_quarter_cache: Dict[Tuple[date, str], Dict[str, Any]] = {}

    def _quarter_shift(quarter_end_in: Any, delta_quarters: int) -> Optional[date]:
        if not isinstance(quarter_end_in, date):
            return None
        try:
            return (pd.Timestamp(quarter_end_in).to_period("Q") + int(delta_quarters)).end_time.normalize().date()
        except Exception:
            return None

    def _coproduct_source_bucket(source_type_in: Any) -> str:
        low = str(source_type_in or "").strip().lower()
        if low.startswith("nwer"):
            return "nwer"
        if low.startswith("ams_3618"):
            return "ams_3618"
        return ""

    def _coproduct_source_label(source_bucket_in: Any) -> str:
        source_bucket_txt = str(source_bucket_in or "").strip().lower()
        return (
            "NWER" if source_bucket_txt == "nwer"
            else "AMS 3618" if source_bucket_txt == "ams_3618"
            else ""
        )

    def _coproduct_series_candidates(input_key: str, region_in: Any) -> Tuple[str, ...]:
        region_txt = str(region_in or "").strip().lower()
        return tuple((coproduct_region_series_candidates.get(str(input_key or "").strip()) or {}).get(region_txt) or ())

    def _coproduct_target_unit(input_key: str) -> str:
        tpl = market_input_templates_by_key.get(str(input_key or "").strip())
        return str(getattr(tpl, "unit", "") or "").strip()

    def _coproduct_best_aggregate_candidate(
        rows_in: List[Dict[str, Any]],
        *,
        target_unit: str,
        agg_preference: str,
    ) -> Optional[Dict[str, Any]]:
        def _best_non_obs(agg_level_txt: str) -> Optional[Dict[str, Any]]:
            best_score: Optional[Tuple[Any, ...]] = None
            best_rec: Optional[Dict[str, Any]] = None
            for rec in rows_in:
                if str(rec.get("aggregation_level") or "").strip().lower() != agg_level_txt:
                    continue
                converted_val, converted = _convert_market_price_value(
                    rec.get("price_value"),
                    str(rec.get("unit") or ""),
                    target_unit,
                )
                if converted_val is None:
                    continue
                score = (
                    _market_quality_rank(rec.get("quality")),
                    _safe_int_like(rec.get("_obs_count")),
                    _safe_int_like(rec.get("observation_date_ord")),
                )
                if best_score is None or score > best_score:
                    chosen = dict(rec)
                    chosen["_converted_value"] = float(converted_val)
                    chosen["_converted"] = bool(converted)
                    chosen["_selection_rule"] = agg_level_txt
                    best_score = score
                    best_rec = chosen
            return best_rec

        for agg_level_txt in (str(agg_preference or "").strip().lower() or "quarter_avg", "quarter_end"):
            chosen = _best_non_obs(agg_level_txt)
            if isinstance(chosen, dict):
                return chosen
        observation_vals: List[float] = []
        representative_rec: Optional[Dict[str, Any]] = None
        for rec in rows_in:
            if str(rec.get("aggregation_level") or "").strip().lower() != "observation":
                continue
            converted_val, converted = _convert_market_price_value(
                rec.get("price_value"),
                str(rec.get("unit") or ""),
                target_unit,
            )
            if converted_val is None:
                continue
            observation_vals.append(float(converted_val))
            if representative_rec is None:
                representative_rec = dict(rec)
                representative_rec["_converted"] = bool(converted)
        if observation_vals and representative_rec is not None:
            representative_rec["_converted_value"] = float(sum(observation_vals) / len(observation_vals))
            representative_rec["_selection_rule"] = "observation_avg"
            representative_rec["_obs_count"] = len(observation_vals)
            return representative_rec
        return None

    def _coproduct_quarter_open_candidate(
        rows_in: List[Dict[str, Any]],
        *,
        target_unit: str,
    ) -> Optional[Dict[str, Any]]:
        earliest_by_day: Dict[date, List[Tuple[Tuple[Any, ...], Dict[str, Any]]]] = {}
        for rec in rows_in:
            if str(rec.get("aggregation_level") or "").strip().lower() != "observation":
                continue
            obs_dt = pd.to_datetime(rec.get("observation_date"), errors="coerce")
            if pd.isna(obs_dt):
                continue
            converted_val, converted = _convert_market_price_value(
                rec.get("price_value"),
                str(rec.get("unit") or ""),
                target_unit,
            )
            if converted_val is None:
                continue
            obs_date = pd.Timestamp(obs_dt).date()
            score = (
                -_market_quality_rank(rec.get("quality")),
                -_safe_int_like(rec.get("_obs_count")),
            )
            chosen = dict(rec)
            chosen["_converted_value"] = float(converted_val)
            chosen["_converted"] = bool(converted)
            chosen["_selection_rule"] = "quarter_open_observation"
            earliest_by_day.setdefault(obs_date, []).append((score, chosen))
        if not earliest_by_day:
            return None
        first_obs_date = min(earliest_by_day)
        day_candidates = sorted(earliest_by_day.get(first_obs_date) or [], key=lambda item: item[0])
        if not day_candidates:
            return None
        return dict(day_candidates[0][1])

    def _resolved_uhp_value_for_quarter(quarter_end: date, *, mode: str) -> Dict[str, Any]:
        tpl = market_input_templates_by_key.get("uhp_price")
        if tpl is None or not isinstance(quarter_end, date):
            return {"value": None, "source_mode": "Unknown/blank"}
        target_unit = str(getattr(tpl, "unit", "") or "").strip()
        series_keys = tuple(
            str(item or "").strip()
            for item in (getattr(tpl, "source_series_keys", ()) or ())
            if str(item or "").strip()
        )
        candidates: List[Tuple[Tuple[Any, ...], Dict[str, Any]]] = []
        for rec in economics_market_rows:
            if rec.get("quarter") != quarter_end or str(rec.get("series_key") or "").strip() not in series_keys:
                continue
            source_bucket = _coproduct_source_bucket(rec.get("source_type"))
            source_rank = coproduct_source_priority.index(source_bucket) if source_bucket in coproduct_source_priority else len(coproduct_source_priority)
            if str(mode or "").strip().lower() == "quarter_open":
                chosen = _coproduct_quarter_open_candidate([rec], target_unit=target_unit)
            else:
                chosen = _coproduct_best_aggregate_candidate(
                    [rec],
                    target_unit=target_unit,
                    agg_preference=str(getattr(tpl, "aggregation_preference", "") or "quarter_avg"),
                )
            if not isinstance(chosen, dict):
                continue
            candidates.append(((source_rank,), chosen))
        if not candidates:
            return {"value": None, "source_mode": "Unknown/blank"}
        picked = sorted(candidates, key=lambda item: item[0])[0][1]
        value_num = pd.to_numeric(picked.get("_converted_value"), errors="coerce")
        return {
            "value": None if pd.isna(value_num) else float(value_num),
            "source_mode": _coproduct_source_label(_coproduct_source_bucket(picked.get("source_type"))) or "Unknown/blank",
        }

    def _weighted_coproduct_input_value(
        input_key: str,
        quarter_end: date,
        *,
        mode: str = "quarter_avg",
    ) -> Dict[str, Any]:
        cache_key = (str(input_key or "").strip(), quarter_end, str(mode or "").strip())
        cached = weighted_coproduct_input_cache.get(cache_key)
        if isinstance(cached, dict):
            return dict(cached)
        if not isinstance(quarter_end, date):
            out = {"value": None, "coverage_ratio": None, "source_mode": "Unknown/blank", "region_hits": []}
            weighted_coproduct_input_cache[cache_key] = dict(out)
            return out
        tpl = market_input_templates_by_key.get(str(input_key or "").strip())
        if tpl is None:
            out = {"value": None, "coverage_ratio": None, "source_mode": "Unknown/blank", "region_hits": []}
            weighted_coproduct_input_cache[cache_key] = dict(out)
            return out
        target_unit = _coproduct_target_unit(input_key)
        agg_preference = str(getattr(tpl, "aggregation_preference", "") or "quarter_avg").strip().lower()
        weights = {
            str(region or "").strip().lower(): float(pd.to_numeric(raw_weight, errors="coerce") or 0.0)
            for region, raw_weight in dict(
                _gpre_official_market_weights_for_quarter(
                    quarter_end,
                    ticker_root=gpre_ticker_root_local,
                    plant_capacity_history=gpre_plant_capacity_history,
                )
                or {}
            ).items()
            if float(pd.to_numeric(raw_weight, errors="coerce") or 0.0) > 0.0
        }
        total_weight = float(sum(weights.values()))
        if total_weight <= 0.0:
            out = {"value": None, "coverage_ratio": None, "source_mode": "Unknown/blank", "region_hits": []}
            weighted_coproduct_input_cache[cache_key] = dict(out)
            return out
        covered_weight = 0.0
        weighted_total = 0.0
        source_labels: List[str] = []
        region_hits: List[str] = []
        for region, region_weight in sorted(weights.items()):
            candidate: Optional[Dict[str, Any]] = None
            for source_bucket in coproduct_source_priority:
                for series_key in _coproduct_series_candidates(input_key, region):
                    matching_rows = [
                        rec
                        for rec in economics_market_rows
                        if rec.get("quarter") == quarter_end
                        and str(rec.get("series_key") or "").strip() == series_key
                        and _coproduct_source_bucket(rec.get("source_type")) == source_bucket
                    ]
                    if not matching_rows:
                        continue
                    if str(mode or "").strip().lower() == "quarter_open":
                        chosen = _coproduct_quarter_open_candidate(matching_rows, target_unit=target_unit)
                    else:
                        chosen = _coproduct_best_aggregate_candidate(
                            matching_rows,
                            target_unit=target_unit,
                            agg_preference=agg_preference,
                        )
                    if isinstance(chosen, dict):
                        candidate = dict(chosen)
                        candidate["_series_key"] = series_key
                        candidate["_source_bucket"] = source_bucket
                        break
                if candidate is not None:
                    break
            if candidate is None:
                continue
            value_num = pd.to_numeric(candidate.get("_converted_value"), errors="coerce")
            if pd.isna(value_num):
                continue
            covered_weight += float(region_weight)
            weighted_total += float(region_weight) * float(value_num)
            source_label = _coproduct_source_label(candidate.get("_source_bucket"))
            if source_label:
                source_labels.append(source_label)
            region_hits.append(f"{region}:{candidate.get('_series_key')}:{source_label or 'Unknown'}")
        coverage_ratio = (covered_weight / total_weight) if total_weight > 0 else None
        value_out = (weighted_total / covered_weight) if covered_weight > 0 else None
        out = {
            "value": value_out,
            "coverage_ratio": coverage_ratio,
            "source_mode": _classify_coproduct_resolved_source(*source_labels),
            "region_hits": list(region_hits),
        }
        weighted_coproduct_input_cache[cache_key] = dict(out)
        return out

    def _derive_coproduct_credit_record(
        *,
        quarter_end: Optional[date],
        quarter_label: str,
        corn_oil_price: Optional[float],
        distillers_price: Optional[float],
        uhp_price: Optional[float],
        corn_oil_source_mode: str,
        distillers_source_mode: str,
        corn_oil_coverage_ratio: Optional[float],
        distillers_coverage_ratio: Optional[float],
        gallons_million_display: Optional[float],
        rule_text: str,
        yield_anchor: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        yield_anchor = dict(yield_anchor or {})
        distillers_yield_used = pd.to_numeric(yield_anchor.get("distillers_yield_lbs_per_bu"), errors="coerce")
        if pd.isna(distillers_yield_used):
            distillers_yield_used = distillers_yield_num
        uhp_yield_used = pd.to_numeric(yield_anchor.get("uhp_yield_lbs_per_bu"), errors="coerce")
        if pd.isna(uhp_yield_used):
            uhp_yield_used = uhp_yield_num
        corn_oil_yield_used = pd.to_numeric(yield_anchor.get("renewable_corn_oil_yield_lbs_per_bu"), errors="coerce")
        if pd.isna(corn_oil_yield_used):
            corn_oil_yield_used = corn_oil_yield_num
        distillers_contribution = (
            float(distillers_yield_used) * float(distillers_price)
            if pd.notna(distillers_yield_used) and distillers_price is not None
            else None
        )
        uhp_contribution = (
            float(uhp_yield_used) * float(uhp_price)
            if pd.notna(uhp_yield_used) and uhp_price is not None
            else None
        )
        corn_oil_contribution = (
            float(corn_oil_yield_used) * float(corn_oil_price)
            if pd.notna(corn_oil_yield_used) and corn_oil_price is not None
            else None
        )
        contribution_values = [
            float(val)
            for val in (distillers_contribution, uhp_contribution, corn_oil_contribution)
            if val is not None and pd.notna(val)
        ]
        approximate_coproduct_credit = sum(contribution_values) if contribution_values else None
        approximate_coproduct_credit_per_gal = (
            float(approximate_coproduct_credit) / float(ethanol_yield_num)
            if approximate_coproduct_credit is not None
            and pd.notna(ethanol_yield_num)
            and abs(float(ethanol_yield_num)) > 1e-9
            else None
        )
        renewable_corn_oil_contribution_per_gal = (
            float(corn_oil_contribution) / float(ethanol_yield_num)
            if corn_oil_contribution is not None
            and pd.notna(ethanol_yield_num)
            and abs(float(ethanol_yield_num)) > 1e-9
            else None
        )
        gallons_num = pd.to_numeric(gallons_million_display, errors="coerce")
        approximate_coproduct_credit_usd_m = (
            float(approximate_coproduct_credit_per_gal) * float(gallons_num)
            if approximate_coproduct_credit_per_gal is not None
            and pd.notna(gallons_num)
            and abs(float(gallons_num)) > 1e-9
            else None
        )
        renewable_corn_oil_contribution_usd_m_proxy = (
            float(renewable_corn_oil_contribution_per_gal) * float(gallons_num)
            if renewable_corn_oil_contribution_per_gal is not None
            and pd.notna(gallons_num)
            and abs(float(gallons_num)) > 1e-9
            else None
        )
        coverage_candidates = [
            float(val)
            for val in (corn_oil_coverage_ratio, distillers_coverage_ratio)
            if pd.notna(pd.to_numeric(val, errors="coerce"))
        ]
        coverage_ratio = min(coverage_candidates) if coverage_candidates else None
        return {
            "quarter_end": quarter_end,
            "quarter_label": quarter_label,
            "renewable_corn_oil_price": corn_oil_price,
            "distillers_grains_price": distillers_price,
            "uhp_price": uhp_price,
            "renewable_corn_oil_contribution_per_bushel": corn_oil_contribution,
            "renewable_corn_oil_contribution_per_gal": renewable_corn_oil_contribution_per_gal,
            "renewable_corn_oil_contribution_usd_m_proxy": renewable_corn_oil_contribution_usd_m_proxy,
            "approximate_coproduct_credit": approximate_coproduct_credit,
            "approximate_coproduct_credit_per_gal": approximate_coproduct_credit_per_gal,
            "approximate_coproduct_credit_usd_m": approximate_coproduct_credit_usd_m,
            "resolved_source_mode": _classify_coproduct_resolved_source(
                corn_oil_source_mode,
                distillers_source_mode,
            ),
            "coverage_ratio": coverage_ratio,
            "rule": rule_text,
            "volume_yield_source_mode": str(yield_anchor.get("source_mode") or ""),
            "volume_yield_anchor_quarter": yield_anchor.get("anchor_quarter"),
            "volume_yield_anchor_quarter_label": str(yield_anchor.get("anchor_quarter_label") or ""),
            "distillers_yield_used_lbs_per_bu": (None if pd.isna(distillers_yield_used) else float(distillers_yield_used)),
            "uhp_yield_used_lbs_per_bu": (None if pd.isna(uhp_yield_used) else float(uhp_yield_used)),
            "renewable_corn_oil_yield_used_lbs_per_bu": (None if pd.isna(corn_oil_yield_used) else float(corn_oil_yield_used)),
            "renewable_corn_oil_source_mode": corn_oil_source_mode or "Unknown/blank",
            "renewable_corn_oil_price_source_mode": corn_oil_source_mode or "Unknown/blank",
            "distillers_grains_source_mode": distillers_source_mode or "Unknown/blank",
            "distillers_grains_price_source_mode": distillers_source_mode or "Unknown/blank",
            "renewable_corn_oil_coverage_ratio": corn_oil_coverage_ratio,
            "renewable_corn_oil_price_coverage_ratio": corn_oil_coverage_ratio,
            "distillers_grains_coverage_ratio": distillers_coverage_ratio,
            "distillers_grains_price_coverage_ratio": distillers_coverage_ratio,
        }

    def _weighted_coproduct_quarter_record(
        quarter_end: date,
        *,
        mode: str = "quarter_avg",
    ) -> Dict[str, Any]:
        cache_key = (quarter_end, str(mode or "").strip())
        cached = weighted_coproduct_quarter_cache.get(cache_key)
        if isinstance(cached, dict):
            return dict(cached)
        if not isinstance(quarter_end, date):
            return {}
        corn_oil_rec = _weighted_coproduct_input_value("renewable_corn_oil_price", quarter_end, mode=mode)
        distillers_rec = _weighted_coproduct_input_value("distillers_grains_price", quarter_end, mode=mode)
        uhp_rec = _resolved_uhp_value_for_quarter(quarter_end, mode=mode)
        out = _derive_coproduct_credit_record(
            quarter_end=quarter_end,
            quarter_label=_quarter_label_short(quarter_end),
            corn_oil_price=pd.to_numeric(corn_oil_rec.get("value"), errors="coerce"),
            distillers_price=pd.to_numeric(distillers_rec.get("value"), errors="coerce"),
            uhp_price=pd.to_numeric(uhp_rec.get("value"), errors="coerce"),
            corn_oil_source_mode=str(corn_oil_rec.get("source_mode") or "Unknown/blank"),
            distillers_source_mode=str(distillers_rec.get("source_mode") or "Unknown/blank"),
            corn_oil_coverage_ratio=pd.to_numeric(corn_oil_rec.get("coverage_ratio"), errors="coerce"),
            distillers_coverage_ratio=pd.to_numeric(distillers_rec.get("coverage_ratio"), errors="coerce"),
            gallons_million_display=historical_gallons_million_map.get(quarter_end),
            rule_text=(
                "Weighted exact-quarter averages using quarter-aware active-capacity footprint; corn oil uses the all-active-footprint approximation."
                if str(mode or "").strip().lower() != "quarter_open"
                else "Weighted early-quarter observation snapshot using quarter-aware active-capacity footprint; missing legs stay blank for carry-forward handling."
            ),
        )
        weighted_coproduct_quarter_cache[cache_key] = dict(out)
        return out

    def _coproduct_frame_record(
        frame_key: str,
        *,
        target_quarter_end: Optional[date],
        base_record: Optional[Dict[str, Any]],
        fallback_record: Optional[Dict[str, Any]] = None,
        gallons_million_display: Optional[float] = None,
    ) -> Dict[str, Any]:
        frame_label = {
            "prior_quarter": "Prior quarter",
            "quarter_open": "Quarter-open outlook",
            "current_qtd": "Current QTD",
            "next_quarter_thesis": "Next quarter outlook",
        }.get(str(frame_key or "").strip(), str(frame_key or "").replace("_", " ").title())
        primary = dict(base_record or {})
        fallback = dict(fallback_record or {})

        def _choose(field_name: str) -> Tuple[Optional[float], str, Optional[float], bool]:
            primary_val = pd.to_numeric(primary.get(field_name), errors="coerce")
            if pd.notna(primary_val):
                return (
                    float(primary_val),
                    str(primary.get(f"{field_name}_source_mode") or primary.get("resolved_source_mode") or "Unknown/blank"),
                    pd.to_numeric(primary.get(f"{field_name}_coverage_ratio") or primary.get("coverage_ratio"), errors="coerce"),
                    False,
                )
            fallback_val = pd.to_numeric(fallback.get(field_name), errors="coerce")
            if pd.notna(fallback_val):
                return (
                    float(fallback_val),
                    str(fallback.get(f"{field_name}_source_mode") or fallback.get("resolved_source_mode") or "Unknown/blank"),
                    pd.to_numeric(fallback.get(f"{field_name}_coverage_ratio") or fallback.get("coverage_ratio"), errors="coerce"),
                    True,
                )
            return None, "Unknown/blank", None, False

        corn_price, corn_source_mode, corn_coverage_ratio, corn_carried = _choose("renewable_corn_oil_price")
        distillers_price, distillers_source_mode, distillers_coverage_ratio, ddgs_carried = _choose("distillers_grains_price")
        uhp_value_num = pd.to_numeric(primary.get("uhp_price"), errors="coerce")
        if pd.isna(uhp_value_num):
            uhp_value_num = pd.to_numeric(fallback.get("uhp_price"), errors="coerce")
        if str(frame_key or "").strip() == "quarter_open":
            rule_text = (
                "Early-quarter weighted observation snapshot where available; prior-quarter carry-forward fills missing coproduct legs."
                if (corn_carried or ddgs_carried)
                else "Early-quarter weighted observation snapshot using the active-capacity footprint."
            )
        elif str(frame_key or "").strip() == "next_quarter_thesis":
            rule_text = "Freeze the resolved quarter-open weighted coproduct frame for next-quarter outlook; fall back to prior quarter only when quarter-open is unavailable."
        elif str(frame_key or "").strip() == "current_qtd":
            rule_text = "Weighted current-quarter averages using the active-capacity footprint."
        else:
            rule_text = "Weighted prior-quarter averages using the active-capacity footprint."
        yield_anchor = _latest_coproduct_yield_anchor(target_quarter_end)
        yield_anchor_note = str(yield_anchor.get("note") or "").strip()
        if yield_anchor_note:
            rule_text = f"{rule_text} {yield_anchor_note}"
        out = _derive_coproduct_credit_record(
            quarter_end=target_quarter_end,
            quarter_label=str(frame_label or ""),
            corn_oil_price=corn_price,
            distillers_price=distillers_price,
            uhp_price=(None if pd.isna(uhp_value_num) else float(uhp_value_num)),
            corn_oil_source_mode=corn_source_mode,
            distillers_source_mode=distillers_source_mode,
            corn_oil_coverage_ratio=corn_coverage_ratio,
            distillers_coverage_ratio=distillers_coverage_ratio,
            gallons_million_display=gallons_million_display,
            rule_text=rule_text,
            yield_anchor=yield_anchor,
        )
        out["frame_key"] = str(frame_key or "")
        out["frame_label"] = frame_label
        out["quarter_end"] = target_quarter_end
        return out

    def _quarterly_coproduct_history_records() -> List[Dict[str, Any]]:
        history_cutoff = current_market_display_quarter if isinstance(current_market_display_quarter, date) else as_of_market_quarter
        if not economics_market_rows or not isinstance(history_cutoff, date):
            return []
        relevant_series_keys: Set[str] = set()
        for input_key in ("renewable_corn_oil_price", "distillers_grains_price", "uhp_price"):
            tpl = market_input_templates_by_key.get(input_key)
            if tpl is None:
                continue
            relevant_series_keys.update(
                str(item or "").strip()
                for item in (getattr(tpl, "source_series_keys", ()) or ())
                if str(item or "").strip()
            )
        if not relevant_series_keys:
            return []
        history_quarters = sorted(
            {
                rec_q
                for rec_q in (
                    rec.get("quarter")
                    for rec in economics_market_rows
                    if str(rec.get("series_key") or "").strip() in relevant_series_keys
                )
                if isinstance(rec_q, date) and rec_q <= history_cutoff
            }
        )
        return [_weighted_coproduct_quarter_record(quarter_end) for quarter_end in history_quarters]

    def _coproduct_frame_summary_records(
        history_records_in: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        history_map = {
            rec.get("quarter_end"): dict(rec)
            for rec in list(history_records_in or [])
            if isinstance(rec, dict) and isinstance(rec.get("quarter_end"), date)
        }
        current_q = current_market_display_quarter if isinstance(current_market_display_quarter, date) else None
        prior_q = (
            prior_market_display_quarter
            if isinstance(prior_market_display_quarter, date)
            else _quarter_shift(current_q, -1)
        )
        quarter_open_q = quarter_open_display_quarter if isinstance(quarter_open_display_quarter, date) else current_q
        next_q = next_thesis_quarter_end if isinstance(next_thesis_quarter_end, date) else _quarter_shift(current_q, 1)
        prior_record = history_map.get(prior_q) or (_weighted_coproduct_quarter_record(prior_q) if isinstance(prior_q, date) else {})
        current_record = history_map.get(current_q) or (_weighted_coproduct_quarter_record(current_q) if isinstance(current_q, date) else {})
        open_base_record = _weighted_coproduct_quarter_record(quarter_open_q, mode="quarter_open") if isinstance(quarter_open_q, date) else {}
        next_quarter_seed_record = open_base_record if any(
            pd.notna(pd.to_numeric((open_base_record or {}).get(field_name), errors="coerce"))
            for field_name in ("renewable_corn_oil_price", "distillers_grains_price", "approximate_coproduct_credit_per_gal")
        ) else prior_record
        return [
            _coproduct_frame_record(
                "prior_quarter",
                target_quarter_end=prior_q,
                base_record=prior_record,
                gallons_million_display=pd.to_numeric(
                    (_gpre_proxy_implied_frame_record("prior_quarter") or {}).get("implied_gallons_million_display"),
                    errors="coerce",
                ),
            ),
            _coproduct_frame_record(
                "quarter_open",
                target_quarter_end=quarter_open_q,
                base_record=open_base_record,
                fallback_record=prior_record,
                gallons_million_display=pd.to_numeric(
                    (_gpre_proxy_implied_frame_record("quarter_open") or {}).get("implied_gallons_million_display"),
                    errors="coerce",
                ),
            ),
            _coproduct_frame_record(
                "current_qtd",
                target_quarter_end=current_q,
                base_record=current_record,
                fallback_record=prior_record,
                gallons_million_display=pd.to_numeric(
                    (_gpre_proxy_implied_frame_record("current_qtd") or {}).get("implied_gallons_million_display"),
                    errors="coerce",
                ),
            ),
            _coproduct_frame_record(
                "next_quarter_thesis",
                target_quarter_end=next_q,
                base_record=next_quarter_seed_record,
                fallback_record=prior_record,
                gallons_million_display=pd.to_numeric(
                    (_gpre_proxy_implied_frame_record("next_quarter_thesis") or {}).get("implied_gallons_million_display"),
                    errors="coerce",
                ),
            ),
        ]

    def _coproduct_filled_now_text(state_in: Dict[str, bool]) -> str:
        bits: List[str] = []
        if bool((state_in or {}).get("historical")):
            bits.append("Hist")
        if bool((state_in or {}).get("current")):
            bits.append("Current")
        if bool((state_in or {}).get("next")):
            bits.append("Next")
        return " + ".join(bits) if bits else "Blank"

    def _coproduct_readiness_bucket_text(state_in: Dict[str, bool], bucket: str, readiness_kind: str) -> str:
        ready = bool((state_in or {}).get(bucket))
        if readiness_kind == "assumption":
            return "Assumption" if ready else "Missing assumption"
        if readiness_kind == "derived":
            return "Ready" if ready else "Blocked"
        return "Ready" if ready else "Needs source"

    def _metrics_pick(model_key_in: str, split_in: str, field_in: str) -> Optional[float]:
        if not isinstance(metrics_df, pd.DataFrame) or metrics_df.empty:
            return None
        sub = metrics_df[
            (metrics_df["model_key"].astype(str) == str(model_key_in or ""))
            & (metrics_df["split"].astype(str) == str(split_in or ""))
        ].copy()
        if sub.empty:
            return None
        return pd.to_numeric(sub.iloc[0].get(field_in), errors="coerce")

    bridge_official_key = str(model_result.get("gpre_proxy_model_key") or model_result.get("bridge_official_model_key") or "process_current_quarter_avg")
    recommended_display = _sandbox_model_label(bridge_official_key)
    incumbent_display = _sandbox_model_label(incumbent_baseline_model_key)
    process_comparator_display = _sandbox_model_label("process_front_loaded")
    expanded_candidate_display = _sandbox_model_label(expanded_best_candidate_model_key)
    best_historical_fit_display = _sandbox_model_label(best_historical_fit_model_key)
    best_compromise_display = _sandbox_model_label(best_compromise_model_key)
    best_forward_lens_display = _sandbox_model_label(best_forward_lens_model_key)
    production_winner_display = _sandbox_model_label(production_winner_model_key)
    simple_market_test_mae = _metrics_pick("simple_market", "test", "mae")
    recommended_test_mae = _metrics_pick(bridge_official_key, "test", "mae")
    recommended_test_corr = _metrics_pick(bridge_official_key, "test", "correlation")
    chosen_row = {}
    expanded_best_row = {}
    best_historical_fit_row = {}
    best_compromise_row = {}
    best_forward_lens_row = {}
    if isinstance(leaderboard_df, pd.DataFrame) and not leaderboard_df.empty:
        chosen_sub = leaderboard_df[leaderboard_df["chosen"] == True].copy()
        if not chosen_sub.empty:
            chosen_row = chosen_sub.iloc[0].to_dict()
        expanded_sub = leaderboard_df[leaderboard_df["expanded_best_candidate"] == True].copy() if "expanded_best_candidate" in leaderboard_df.columns else pd.DataFrame()
        if not expanded_sub.empty:
            expanded_best_row = expanded_sub.iloc[0].to_dict()
        best_historical_sub = leaderboard_df[leaderboard_df["model_key"].astype(str) == best_historical_fit_model_key].copy()
        if not best_historical_sub.empty:
            best_historical_fit_row = best_historical_sub.iloc[0].to_dict()
        best_compromise_sub = leaderboard_df[leaderboard_df["model_key"].astype(str) == best_compromise_model_key].copy()
        if not best_compromise_sub.empty:
            best_compromise_row = best_compromise_sub.iloc[0].to_dict()
        best_forward_sub = leaderboard_df[leaderboard_df["model_key"].astype(str) == best_forward_lens_model_key].copy()
        if not best_forward_sub.empty:
            best_forward_lens_row = best_forward_sub.iloc[0].to_dict()
    chosen_family_txt = str(chosen_row.get("family_label") or "")
    chosen_timing_txt = str(chosen_row.get("timing_rule") or "")
    chosen_clean_mae = pd.to_numeric(chosen_row.get("clean_mae"), errors="coerce")
    chosen_underlying_mae = pd.to_numeric(chosen_row.get("underlying_mae"), errors="coerce")
    chosen_hybrid = pd.to_numeric(chosen_row.get("hybrid_score"), errors="coerce")
    chosen_forward_usability = str(chosen_row.get("forward_usability_rating") or "").strip()
    best_forward_usability = str(best_forward_lens_row.get("forward_usability_rating") or "").strip()
    chosen_preview_quality = str(
        chosen_row.get("live_preview_quality_status")
        or model_result.get("gpre_proxy_live_preview_quality_status")
        or ""
    ).strip()
    chosen_preview_mae = pd.to_numeric(
        chosen_row.get("live_preview_mae")
        if "live_preview_mae" in chosen_row
        else model_result.get("gpre_proxy_live_preview_mae"),
        errors="coerce",
    )
    chosen_preview_max_error = pd.to_numeric(
        chosen_row.get("live_preview_max_error")
        if "live_preview_max_error" in chosen_row
        else model_result.get("gpre_proxy_live_preview_max_error"),
        errors="coerce",
    )
    chosen_preview_top_miss = str(
        chosen_row.get("live_preview_top_miss_quarters")
        or model_result.get("gpre_proxy_live_preview_top_miss_quarters")
        or ""
    ).strip()
    chosen_preview_worst_phase = str(
        chosen_row.get("live_preview_worst_phase")
        or model_result.get("gpre_proxy_live_preview_worst_phase")
        or ""
    ).strip()
    chosen_hard_mae = pd.to_numeric(
        chosen_row.get("hard_quarter_mae")
        if "hard_quarter_mae" in chosen_row
        else model_result.get("gpre_proxy_hard_quarter_mae"),
        errors="coerce",
    )
    chosen_hard_count = pd.to_numeric(
        chosen_row.get("hard_quarter_count")
        if "hard_quarter_count" in chosen_row
        else model_result.get("gpre_proxy_hard_quarter_count"),
        errors="coerce",
    )
    chosen_hard_top_miss = str(
        chosen_row.get("hard_quarter_top_miss_quarters")
        or model_result.get("gpre_proxy_hard_quarter_top_miss_quarters")
        or ""
    ).strip()
    hedge_candidate_leaderboard_df = (
        hedge_style_study.get("candidate_leaderboard_df")
        if isinstance(hedge_style_study.get("candidate_leaderboard_df"), pd.DataFrame)
        else pd.DataFrame()
    )
    hedge_quarter_fit_df = (
        hedge_style_study.get("quarter_fit_df")
        if isinstance(hedge_style_study.get("quarter_fit_df"), pd.DataFrame)
        else pd.DataFrame()
    )
    hedge_backtest_window_display = str(hedge_style_study.get("backtest_window_display") or "").strip()
    hedge_target_label = str(hedge_style_study.get("target_label") or "Reported consolidated crush margin ($/gal)")
    hedge_target_definition = str(hedge_style_study.get("target_definition") or "").strip()
    hedge_best_style_display = _sandbox_model_label(hedge_style_study.get("best_overall_style_key") or hedge_style_study.get("best_overall_style_label") or "")
    hedge_best_family_display = str(hedge_style_study.get("best_overall_style_family_label") or hedge_style_study.get("best_overall_style_family") or "").strip()
    hedge_style_vs_family_explanation = str(hedge_style_study.get("best_style_vs_family_explanation") or "").strip()
    hedge_diagnostic_only_note = str(hedge_style_study.get("diagnostic_only_note") or "").strip()
    hedge_interpretation_lines = [
        str(item or "").strip()
        for item in list(hedge_style_study.get("interpretation_lines") or [])
        if str(item or "").strip()
    ]
    futures_timing_study = (
        model_result.get("futures_timing_study")
        if isinstance(model_result.get("futures_timing_study"), dict)
        else {}
    )
    futures_timing_leaderboard_df = (
        futures_timing_study.get("candidate_leaderboard_df")
        if isinstance(futures_timing_study.get("candidate_leaderboard_df"), pd.DataFrame)
        else pd.DataFrame()
    )
    futures_timing_detail_df = (
        futures_timing_study.get("quarter_detail_df")
        if isinstance(futures_timing_study.get("quarter_detail_df"), pd.DataFrame)
        else pd.DataFrame()
    )
    futures_timing_target_label = str(
        futures_timing_study.get("target_label")
        or "Evaluation target crush margin ($/gal)"
    ).strip()
    futures_timing_coverage_note = str(futures_timing_study.get("coverage_note") or "").strip()
    futures_timing_diagnostic_note = str(futures_timing_study.get("diagnostic_only_note") or "").strip()
    system_audit = model_result.get("system_audit") if isinstance(model_result.get("system_audit"), dict) else {}
    system_official_row_role = str(system_audit.get("official_row_role") or "Approximate market crush = simple market/process proxy").strip()
    system_fitted_row_role = str(system_audit.get("fitted_row_role") or "GPRE crush proxy = fitted production model").strip()
    system_expanded_pass_role = str(system_audit.get("expanded_pass_role") or "Expanded-pass best = best challenger in the expanded test set").strip()
    system_production_winner_role = str(system_audit.get("production_winner_role") or "Production winner = model that cleared promotion guardrails").strip()
    system_best_historical_role = str(system_audit.get("best_historical_fit_role") or "Best historical fit = lowest clean-window MAE among eligible official rows").strip()
    system_best_compromise_role = str(system_audit.get("best_compromise_role") or "Best compromise = best preview-supported fit/robustness balance").strip()
    system_best_forward_role = str(system_audit.get("best_forward_lens_role") or "Best forward lens = strongest forward-usable preview lens").strip()
    system_winner_preview_quality = str(system_audit.get("winner_preview_quality") or chosen_preview_quality or "n/a").strip()
    system_winner_forward_usability = str(system_audit.get("winner_forward_usability") or chosen_forward_usability or "n/a").strip()
    system_hedge_role = str(system_audit.get("hedge_style_study_role") or hedge_diagnostic_only_note or "Diagnostic only; does not change official row, fitted row, or winner selection.").strip()
    system_internal_inconsistency = bool(system_audit.get("internal_consistency_detected"))
    recent_quarter_comparison_df = (
        model_result.get("recent_quarter_comparison_df")
        if isinstance(model_result.get("recent_quarter_comparison_df"), pd.DataFrame)
        else pd.DataFrame()
    )
    expanded_best_selection_txt = (
        _sandbox_guard_label(expanded_best_row.get("selection_guard_reason"))
        if expanded_best_row
        else ""
    )
    expanded_best_promotion_txt = (
        _sandbox_guard_label(expanded_best_row.get("promotion_guard_reason"))
        if expanded_best_row
        else ""
    )
    expanded_preview_quality = str(expanded_best_row.get("live_preview_quality_status") or "").strip()
    expanded_preview_mae = pd.to_numeric(expanded_best_row.get("live_preview_mae"), errors="coerce")
    expanded_preview_max_error = pd.to_numeric(expanded_best_row.get("live_preview_max_error"), errors="coerce")
    expanded_preview_worst_phase = str(expanded_best_row.get("live_preview_worst_phase") or "").strip()
    expanded_preview_top_miss = str(expanded_best_row.get("live_preview_top_miss_quarters") or "").strip()
    expanded_promotion_failures = str(expanded_best_row.get("promotion_guard_failures") or "").strip()
    expanded_preview_block_reason = ""
    if (
        expanded_best_row
        and expanded_best_candidate_model_key
        and production_winner_model_key
        and str(expanded_best_candidate_model_key) != str(production_winner_model_key)
        and "live_preview_quality_not_faithful_enough" in expanded_promotion_failures
    ):
        block_bits = [f"{expanded_candidate_display} was blocked by preview fidelity."]
        if expanded_preview_quality:
            block_bits.append(f"Quality {expanded_preview_quality}.")
        if pd.notna(expanded_preview_mae):
            block_bits.append(f"MAE {float(expanded_preview_mae):.3f} $/gal.")
        if pd.notna(expanded_preview_max_error):
            block_bits.append(f"Max error {float(expanded_preview_max_error):.3f} $/gal.")
        if expanded_preview_worst_phase:
            block_bits.append(
                f"Main failing mode {_sandbox_preview_phase_label(expanded_preview_worst_phase)}."
            )
        if expanded_preview_top_miss:
            block_bits.append(f"Worst preview misses: {expanded_preview_top_miss}.")
        expanded_preview_block_reason = " ".join(block_bits).strip()

    target_ws.freeze_panes = "B5"
    width_map = {
        2: 16.5,
        3: 10.5,
        4: 24.0,
        5: 19.0,
        6: 25.0,
        7: 19.0,
        8: 4.5,
        9: 21.0,
        10: 10.0,
        11: 10.5,
        12: 10.5,
        13: 10.5,
        14: 10.5,
        15: 12.5,
        16: 9.0,
        17: 9.0,
        18: 10.0,
        19: 36.0,
        20: 5.0,
        21: 21.0,
        22: 21.0,
        23: 21.0,
        24: 14.0,
    }
    for col_idx, width in width_map.items():
        target_ws.column_dimensions[get_column_letter(col_idx)].width = width

    primary_section_fill = PatternFill("solid", fgColor="D9E7F3")
    secondary_section_fill = PatternFill("solid", fgColor="EAF3FB")
    diagnostic_section_fill = PatternFill("solid", fgColor="F7F9FC")
    primary_header_fill = PatternFill("solid", fgColor="E2EEF9")
    secondary_header_fill = PatternFill("solid", fgColor="EEF5FB")
    note_box_fill = PatternFill("solid", fgColor="F4F8FC")
    diagnostic_note_fill = PatternFill("solid", fgColor="F8FBFD")
    focus_fill = PatternFill("solid", fgColor="E8F1FB")
    spacer_fill = PatternFill(fill_type=None)
    section_title_font = Font(bold=True, size=header_size, color=dark_text_color)
    note_font = Font(size=font_size, color=muted_text_color)
    note_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    frame_text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    frame_value_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    numeric_alignment = Alignment(horizontal="right", vertical="center")
    numeric_wrap_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    medium_side = Side(style="medium", color=border_color)

    def _style_row_range(
        row_idx: int,
        start_col: int,
        end_col: int,
        *,
        fill: Optional[PatternFill] = None,
        font: Optional[Font] = None,
        alignment: Optional[Alignment] = None,
        border: Optional[Border] = None,
        height: Optional[float] = None,
    ) -> None:
        for cc in range(start_col, end_col + 1):
            cell = target_ws.cell(row=row_idx, column=cc)
            if fill is not None:
                cell.fill = copy(fill)
            if font is not None:
                cell.font = copy(font)
            if alignment is not None:
                cell.alignment = copy(alignment)
            if border is not None:
                cell.border = copy(border)
        if height is not None:
            target_ws.row_dimensions[row_idx].height = height

    def _style_box_range(
        row_start: int,
        row_end: int,
        col_start: int,
        col_end: int,
        *,
        fill: Optional[PatternFill] = None,
        font: Optional[Font] = None,
        alignment: Optional[Alignment] = None,
        border: Optional[Border] = None,
        row_height: Optional[float] = None,
    ) -> None:
        for rr in range(row_start, row_end + 1):
            _style_row_range(
                rr,
                col_start,
                col_end,
                fill=fill,
                font=font,
                alignment=alignment,
                border=border,
                height=row_height,
            )

    def _apply_outer_border(row_start: int, row_end: int, col_start: int, col_end: int) -> None:
        for rr in range(row_start, row_end + 1):
            for cc in range(col_start, col_end + 1):
                cell = target_ws.cell(row=rr, column=cc)
                current_border = cell.border
                cell.border = Border(
                    left=copy(medium_side if cc == col_start else current_border.left),
                    right=copy(medium_side if cc == col_end else current_border.right),
                    top=copy(medium_side if rr == row_start else current_border.top),
                    bottom=copy(medium_side if rr == row_end else current_border.bottom),
                )

    def _style_section_title_row(
        row_idx: int,
        start_col: int,
        end_col: int,
        *,
        fill: PatternFill,
        font: Optional[Font] = None,
        height: float = 24.0,
    ) -> None:
        _style_row_range(
            row_idx,
            start_col,
            end_col,
            fill=fill,
            font=(font or section_title_font),
            alignment=align_center,
            border=Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side),
            height=height,
        )

    def _style_note_box_row(
        row_idx: int,
        start_col: int,
        end_col: int,
        *,
        fill: Optional[PatternFill] = None,
        height: Optional[float] = None,
    ) -> None:
        _style_row_range(
            row_idx,
            start_col,
            end_col,
            fill=(fill or note_box_fill),
            font=note_font,
            alignment=note_alignment,
            border=thin_border,
            height=height,
        )

    target_ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=15)
    title_cell = target_ws.cell(row=start_row, column=2, value="Exploratory GPRE basis proxy sandbox (test)")
    title_cell.fill = copy(analysis_theme["title_fill"])
    title_cell.font = copy(analysis_theme["title_font"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = copy(analysis_theme["thin_border"])
    for cc in range(2, 16):
        target_ws.cell(row=start_row, column=cc).fill = copy(analysis_theme["title_fill"])
        target_ws.cell(row=start_row, column=cc).font = copy(analysis_theme["title_font"])
        target_ws.cell(row=start_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
        target_ws.cell(row=start_row, column=cc).border = copy(analysis_theme["thin_border"])
    target_ws.row_dimensions[start_row].height = 24.0

    intro_row = start_row + 1
    target_ws.merge_cells(start_row=intro_row, start_column=2, end_row=intro_row, end_column=15)
    intro_text = (
        "Exploratory only. Uses real AMS 3617 regional corn basis, local footprint regimes, "
        "and the same quarterly result set that is written to the GPRE sidecar files. "
        "The official evaluation window runs through 2025-Q1; later underlying quarters stay diagnostic only."
    )
    intro_cell = target_ws.cell(row=intro_row, column=2, value=intro_text)
    intro_cell.fill = copy(intro_fill)
    intro_cell.font = copy(body_font)
    intro_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    intro_cell.border = copy(thin_border)
    for cc in range(2, 16):
        target_ws.cell(row=intro_row, column=cc).fill = copy(intro_fill)
        target_ws.cell(row=intro_row, column=cc).border = copy(thin_border)
    target_ws.row_dimensions[intro_row].height = 42.0

    footprint_section_row = start_row + 3
    target_ws.merge_cells(start_row=footprint_section_row, start_column=2, end_row=footprint_section_row, end_column=7)
    target_ws.cell(row=footprint_section_row, column=2, value="Footprint / regime summary")
    target_ws.merge_cells(start_row=footprint_section_row, start_column=9, end_row=footprint_section_row, end_column=19)
    target_ws.cell(row=footprint_section_row, column=9, value="Hybrid-score leaderboard")
    target_ws.merge_cells(start_row=footprint_section_row, start_column=21, end_row=footprint_section_row, end_column=24)
    target_ws.cell(row=footprint_section_row, column=21, value="How to read this sheet")
    for (col_start, col_end) in ((2, 7), (9, 19), (21, 24)):
        for cc in range(col_start, col_end + 1):
            target_ws.cell(row=footprint_section_row, column=cc).fill = copy(section_fill)
            target_ws.cell(row=footprint_section_row, column=cc).font = copy(bold_font)
            target_ws.cell(row=footprint_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
            target_ws.cell(row=footprint_section_row, column=cc).border = copy(thin_border)
    target_ws.row_dimensions[footprint_section_row].height = 22.0

    help_boxes = [
        (
            5,
            8,
            (
                "Conclusion\n"
                f"Approximate market crush is the simple official market/process row. The chosen GPRE crush proxy is {recommended_display}"
                f"{f' | incumbent baseline: {incumbent_display}.' if incumbent_baseline_model_key else ''}"
                f" Process comparator: {process_comparator_display}."
                f"{f' Expanded-pass best: {expanded_candidate_display}.' if expanded_best_candidate_model_key else ''}"
                f"{f' Production winner: {production_winner_display}.' if production_winner_model_key else ''}"
                f"{f' Best historical fit: {best_historical_fit_display}.' if best_historical_fit_model_key else ''}"
                f"{f' Best compromise: {best_compromise_display}.' if best_compromise_model_key else ''}"
                f"{f' Best forward lens: {best_forward_lens_display}.' if best_forward_lens_model_key else ''}"
                f"{f' ({chosen_family_txt})' if chosen_family_txt else ''}."
                f"{f' Timing rule: {chosen_timing_txt}.' if chosen_timing_txt else ''}"
                f"{f' Clean MAE: {float(chosen_clean_mae):.3f}.' if pd.notna(chosen_clean_mae) else ''}"
                f"{f' Underlying MAE: {float(chosen_underlying_mae):.3f}.' if pd.notna(chosen_underlying_mae) else ''}"
                f"{f' Hybrid score: {float(chosen_hybrid):.3f}.' if pd.notna(chosen_hybrid) else ''}"
                f"{f' Test MAE: {float(recommended_test_mae):.3f}.' if pd.notna(recommended_test_mae) else ''}"
                f"{f' Test corr: {float(recommended_test_corr):.3f}.' if pd.notna(recommended_test_corr) else ''}"
                f"{f' Expanded-best selection: {expanded_best_selection_txt}.' if expanded_best_selection_txt else ''}"
                f"{f' Expanded-best promotion: {expanded_best_promotion_txt}.' if expanded_best_promotion_txt else ''}"
                f"{f' Preview quality: {chosen_preview_quality}.' if chosen_preview_quality else ''}"
                f"{f' Forward usability: {system_winner_forward_usability}.' if system_winner_forward_usability and system_winner_forward_usability != 'n/a' else ''}"
                f"{f' Preview MAE: {float(chosen_preview_mae):.3f}.' if pd.notna(chosen_preview_mae) else ''}"
                f"{f' Preview max error: {float(chosen_preview_max_error):.3f}.' if pd.notna(chosen_preview_max_error) else ''}"
                f"{f' Main preview mode: {_sandbox_preview_phase_label(chosen_preview_worst_phase)}.' if chosen_preview_worst_phase else ''}"
                f"{f' Preview top misses: {chosen_preview_top_miss}.' if chosen_preview_top_miss else ''}"
                f"{f' Expanded-best preview: {expanded_preview_quality}.' if expanded_preview_quality else ''}"
                f"{f' Expanded-best preview MAE: {float(expanded_preview_mae):.3f}.' if pd.notna(expanded_preview_mae) else ''}"
                f"{f' Expanded-best preview max error: {float(expanded_preview_max_error):.3f}.' if pd.notna(expanded_preview_max_error) else ''}"
                f"{f' Expanded-best main preview mode: {_sandbox_preview_phase_label(expanded_preview_worst_phase)}.' if expanded_preview_worst_phase else ''}"
                f"{f' Preview block: {expanded_preview_block_reason}' if expanded_preview_block_reason else ''}"
                f"{f' Hard-quarter MAE: {float(chosen_hard_mae):.3f} across {int(chosen_hard_count)} flagged quarters.' if pd.notna(chosen_hard_mae) and pd.notna(chosen_hard_count) else ''}"
                f"{f' Hard-quarter misses: {chosen_hard_top_miss}.' if chosen_hard_top_miss else ''}"
                f"{f' Promotion: {_sandbox_guard_label(promotion_guard_reason)}.' if promotion_guard_reason else ''}"
                f"{f' Decision story: {production_decision_story}.' if production_decision_story else ''}"
                f"{f' Selection vs promotion: {selection_vs_promotion_explanation}.' if selection_vs_promotion_explanation else ''}"
            ),
        ),
        (
            9,
            12,
            "Official market model\nApproximate market crush uses a weighted ethanol benchmark less delivered corn (CBOT corn + official weighted corn basis) and fixed natural gas burden. Official corn basis prefers dated GPRE plant bids when available; otherwise it falls back to active-capacity-weighted AMS basis using mapped state/regional series and deterministic fallbacks. GPRE crush proxy is the only row allowed to add company-specific timing, quarter-open blend, ops penalty, or ethanol geography logic on top of that simple benchmark.",
        ),
        (
            13,
            16,
            "How to read the target and timing tests\nClean reported quarters run through 2025-Q1. 2025-Q2 to Q4 are tracked separately as underlying diagnostics. Hybrid score = 50% clean-window MAE + 50% underlying-window MAE. Lower is better. Baseline marks the actual incumbent and the explicit process comparator. Status shows the expanded-pass best candidate and the final production winner. The conclusion box spells out when selection and promotion disagree. Realized GPRE crush margin uses reported consolidated before 2025-Q2 and underlying from 2025-Q2 onward.",
        ),
        (
            17,
            20,
            (
                "Interpretation help\nHigher corn basis raises effective corn cost and lowers crush. "
                "Process-family rows include gas. Bridge-family rows exclude gas. "
                "Hedge-memo rows blend a prior-quarter anchor with current spot conditions. "
                "Avg diff / >2c / >5c show whether the fitted row is meaningfully different from the official row. "
                "Preview quality tells you whether the reduced live workbook preview stays close enough to the full fitted winner."
                f"{f' Approximate market crush test MAE: {float(simple_market_test_mae):.3f}.' if pd.notna(simple_market_test_mae) else ''}"
            ),
        ),
    ]
    for box_start, box_end, box_text in help_boxes:
        target_ws.merge_cells(start_row=box_start, start_column=21, end_row=box_end, end_column=24)
        box_cell = target_ws.cell(row=box_start, column=21, value=box_text)
        box_cell.fill = copy(intro_fill)
        box_cell.font = copy(body_font)
        box_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        box_cell.border = copy(thin_border)
        for rr in range(box_start, box_end + 1):
            target_ws.row_dimensions[rr].height = max(float(target_ws.row_dimensions[rr].height or 0.0), 22.0)
            for cc in range(21, 25):
                target_ws.cell(row=rr, column=cc).fill = copy(intro_fill)
                target_ws.cell(row=rr, column=cc).border = copy(thin_border)

    def _role_summary_text(display_label: str, role_row: Dict[str, Any]) -> str:
        hybrid_num = pd.to_numeric((role_row or {}).get("hybrid_score"), errors="coerce")
        clean_mae_num = pd.to_numeric(
            (role_row or {}).get("clean_mae")
            if "clean_mae" in (role_row or {})
            else (role_row or {}).get("clean_window_mae"),
            errors="coerce",
        )
        forward_txt = str((role_row or {}).get("forward_usability_rating") or "").strip() or "n/a"
        hybrid_txt = f"{float(hybrid_num):.4f}" if pd.notna(hybrid_num) else "n/a"
        clean_mae_txt = f"{float(clean_mae_num):.4f}" if pd.notna(clean_mae_num) else "n/a"
        return f"{display_label or 'n/a'} | Hybrid {hybrid_txt} | MAE {clean_mae_txt} | Forward {forward_txt}"

    role_summary_title_row = 21
    target_ws.merge_cells(start_row=role_summary_title_row, start_column=21, end_row=role_summary_title_row, end_column=24)
    role_summary_title = target_ws.cell(row=role_summary_title_row, column=21, value="Role summary")
    role_summary_title.fill = copy(section_fill)
    role_summary_title.font = copy(bold_font)
    role_summary_title.alignment = Alignment(horizontal="center", vertical="center")
    role_summary_title.border = copy(thin_border)
    for cc in range(21, 25):
        target_ws.cell(row=role_summary_title_row, column=cc).fill = copy(section_fill)
        target_ws.cell(row=role_summary_title_row, column=cc).font = copy(bold_font)
        target_ws.cell(row=role_summary_title_row, column=cc).border = copy(thin_border)
        target_ws.cell(row=role_summary_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")

    role_summary_rows = [
        ("Production winner", _role_summary_text(production_winner_display, chosen_row)),
        ("Best historical fit", _role_summary_text(best_historical_fit_display or "n/a", best_historical_fit_row)),
        ("Best compromise", _role_summary_text(best_compromise_display or "n/a", best_compromise_row)),
        ("Best forward lens", _role_summary_text(best_forward_lens_display or "n/a", best_forward_lens_row)),
    ]
    role_summary_row = role_summary_title_row + 1
    for label_txt, value_txt in role_summary_rows:
        target_ws.merge_cells(start_row=role_summary_row, start_column=22, end_row=role_summary_row, end_column=24)
        for cc in range(21, 25):
            cell = target_ws.cell(row=role_summary_row, column=cc)
            cell.fill = copy(zebra_fill_light if ((role_summary_row - role_summary_title_row) % 2) else zebra_fill_dark)
            cell.border = copy(thin_border)
            cell.font = copy(body_font if cc > 21 else bold_font)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        target_ws.cell(row=role_summary_row, column=21, value=label_txt)
        target_ws.cell(row=role_summary_row, column=22, value=value_txt)
        target_ws.row_dimensions[role_summary_row].height = 28.0
        role_summary_row += 1

    role_summary_note_row = role_summary_row
    target_ws.merge_cells(start_row=role_summary_note_row, start_column=21, end_row=role_summary_note_row, end_column=24)
    role_summary_note = target_ws.cell(
        row=role_summary_note_row,
        column=21,
        value="Production winner = fitted row used in production; Best forward lens = preview-oriented future-quarter lens.",
    )
    role_summary_note.fill = copy(intro_fill)
    role_summary_note.font = copy(body_font)
    role_summary_note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    role_summary_note.border = copy(thin_border)
    for cc in range(21, 25):
        target_ws.cell(row=role_summary_note_row, column=cc).fill = copy(intro_fill)
        target_ws.cell(row=role_summary_note_row, column=cc).border = copy(thin_border)
    target_ws.row_dimensions[role_summary_note_row].height = 34.0

    winner_story_title_row = role_summary_note_row + 2
    target_ws.merge_cells(start_row=winner_story_title_row, start_column=21, end_row=winner_story_title_row, end_column=24)
    winner_story_title = target_ws.cell(row=winner_story_title_row, column=21, value="Winner story")
    winner_story_title.fill = copy(section_fill)
    winner_story_title.font = copy(bold_font)
    winner_story_title.alignment = Alignment(horizontal="center", vertical="center")
    winner_story_title.border = copy(thin_border)
    for cc in range(21, 25):
        target_ws.cell(row=winner_story_title_row, column=cc).fill = copy(section_fill)
        target_ws.cell(row=winner_story_title_row, column=cc).font = copy(bold_font)
        target_ws.cell(row=winner_story_title_row, column=cc).border = copy(thin_border)
        target_ws.cell(row=winner_story_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
    winner_story_rows = [
        ("Official reference", "Approximate market crush"),
        ("Incumbent baseline", incumbent_display),
        ("Process comparator", process_comparator_display),
        ("Expanded-pass best", expanded_candidate_display),
        ("Production winner", production_winner_display),
        ("Best historical fit", best_historical_fit_display or "n/a"),
        ("Best compromise", best_compromise_display or "n/a"),
        ("Best forward lens", best_forward_lens_display or "n/a"),
        ("Forward usability", (
            f"Winner {system_winner_forward_usability or 'n/a'}"
            f"{f' | Forward lens {best_forward_usability}' if best_forward_usability else ''}"
        ).strip(" |")),
        ("Selection status", expanded_best_selection_txt or "n/a"),
        ("Promotion status", expanded_best_promotion_txt or "n/a"),
        ("Preview quality", (
            f"{chosen_preview_quality}"
            f"{f' | MAE {float(chosen_preview_mae):.3f}' if pd.notna(chosen_preview_mae) else ''}"
            f"{f' | Max {float(chosen_preview_max_error):.3f}' if pd.notna(chosen_preview_max_error) else ''}"
        ).strip(" |")),
        ("Main preview mode", _sandbox_preview_phase_label(chosen_preview_worst_phase) or "n/a"),
        ("Expanded-best preview", (
            f"{expanded_preview_quality}"
            f"{f' | MAE {float(expanded_preview_mae):.3f}' if pd.notna(expanded_preview_mae) else ''}"
            f"{f' | Max {float(expanded_preview_max_error):.3f}' if pd.notna(expanded_preview_max_error) else ''}"
            f"{f' | Main {_sandbox_preview_phase_label(expanded_preview_worst_phase)}' if expanded_preview_worst_phase else ''}"
        ).strip(" |") or "n/a"),
        ("Preview block reason", expanded_preview_block_reason or "n/a"),
        ("Hard-quarter status", (
            f"MAE {float(chosen_hard_mae):.3f} across {int(chosen_hard_count)}"
            if pd.notna(chosen_hard_mae) and pd.notna(chosen_hard_count)
            else "n/a"
        )),
        ("Decision story", production_decision_story or selection_vs_promotion_explanation or "n/a"),
    ]
    winner_story_row = winner_story_title_row + 1
    for label_txt, value_txt in winner_story_rows:
        target_ws.merge_cells(start_row=winner_story_row, start_column=22, end_row=winner_story_row, end_column=24)
        label_cell = target_ws.cell(row=winner_story_row, column=21, value=label_txt)
        value_cell = target_ws.cell(row=winner_story_row, column=22, value=value_txt)
        for cc in range(21, 25):
            cell = target_ws.cell(row=winner_story_row, column=cc)
            cell.fill = copy(zebra_fill_light if ((winner_story_row - winner_story_title_row) % 2) else zebra_fill_dark)
            cell.border = copy(thin_border)
            cell.font = copy(body_font if cc > 21 else bold_font)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        target_ws.row_dimensions[winner_story_row].height = 28.0 if label_txt != "Decision story" else 46.0
        winner_story_row += 1

    experimental_rows_df = (
        experimental_candidate_comparison_df[
            experimental_candidate_comparison_df["model_key"].astype(str) != str(incumbent_baseline_model_key or "")
        ].copy()
        if isinstance(experimental_candidate_comparison_df, pd.DataFrame) and not experimental_candidate_comparison_df.empty
        else pd.DataFrame()
    )
    experimental_title_row = winner_story_row + 1
    target_ws.merge_cells(start_row=experimental_title_row, start_column=21, end_row=experimental_title_row, end_column=24)
    experimental_title = target_ws.cell(row=experimental_title_row, column=21, value="Experimental realization / regime candidates")
    experimental_title.fill = copy(section_fill)
    experimental_title.font = copy(bold_font)
    experimental_title.alignment = Alignment(horizontal="center", vertical="center")
    experimental_title.border = copy(thin_border)
    for cc in range(21, 25):
        target_ws.cell(row=experimental_title_row, column=cc).fill = copy(section_fill)
        target_ws.cell(row=experimental_title_row, column=cc).font = copy(bold_font)
        target_ws.cell(row=experimental_title_row, column=cc).border = copy(thin_border)
        target_ws.cell(row=experimental_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
    experimental_row = experimental_title_row + 1
    experimental_best_display = "n/a"
    experimental_promoted_display = "No"
    if not experimental_rows_df.empty:
        experimental_sorted = experimental_rows_df.sort_values(["hybrid_score", "clean_window_mae"], na_position="last")
        experimental_best_key = str(experimental_sorted.iloc[0].get("model_key") or "")
        experimental_best_display = _sandbox_model_label(experimental_best_key)
        experimental_promoted_display = "Yes" if experimental_best_key == str(production_winner_model_key or "") else "No"
    experimental_summary_rows = [
        ("Current winner", incumbent_display),
        ("Best experimental", experimental_best_display),
        ("Promoted?", experimental_promoted_display),
    ]
    for label_txt, value_txt in experimental_summary_rows:
        target_ws.merge_cells(start_row=experimental_row, start_column=22, end_row=experimental_row, end_column=24)
        for cc in range(21, 25):
            cell = target_ws.cell(
                row=experimental_row,
                column=cc,
                value=label_txt if cc == 21 else value_txt if cc == 22 else None,
            )
            cell.fill = copy(zebra_fill_light if ((experimental_row - experimental_title_row) % 2) else zebra_fill_dark)
            cell.border = copy(thin_border)
            cell.font = copy(body_font if cc > 21 else bold_font)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        target_ws.row_dimensions[experimental_row].height = 26.0
        experimental_row += 1
    if not experimental_rows_df.empty:
        for rec in experimental_rows_df.sort_values(["hybrid_score", "clean_window_mae"], na_position="last").to_dict("records"):
            model_label = _sandbox_model_label(rec.get("model_key"))
            value_txt = (
                f"Hybrid Δ {float(pd.to_numeric(rec.get('hybrid_score_delta_vs_incumbent'), errors='coerce')):.3f} | "
                f"Clean Δ {float(pd.to_numeric(rec.get('clean_window_mae_delta_vs_incumbent'), errors='coerce')):.3f} | "
                f"Hard Δ {float(pd.to_numeric(rec.get('hard_quarter_mae_delta_vs_incumbent'), errors='coerce')):.3f} | "
                f"Preview {str(rec.get('preview_quality_class') or 'n/a')} | "
                f"Reason {str(rec.get('promotion_reason_human') or 'n/a')} | "
                f"Improved {str(rec.get('top_improved_quarters_vs_incumbent') or 'n/a')} | "
                f"Worsened {str(rec.get('top_worsened_quarters_vs_incumbent') or 'n/a')}"
            )
            value_txt = str(value_txt or "").replace("Î”", "delta").replace("Δ", "delta")
            value_txt = (
                str(value_txt or "")
                .replace("ÃŽâ€", "delta")
                .replace("Î”", "delta")
                .replace("Δ", "delta")
                .replace("Hybrid delta ", "Hybrid delta vs current winner ")
                .replace("Clean delta ", "Clean delta vs current winner ")
                .replace("Hard delta ", "Hard delta vs current winner ")
            )
            target_ws.merge_cells(start_row=experimental_row, start_column=22, end_row=experimental_row, end_column=24)
            for cc in range(21, 25):
                cell = target_ws.cell(
                    row=experimental_row,
                    column=cc,
                    value=model_label if cc == 21 else value_txt if cc == 22 else None,
                )
                cell.fill = copy(zebra_fill_light if ((experimental_row - experimental_title_row) % 2) else zebra_fill_dark)
                cell.border = copy(thin_border)
                cell.font = copy(body_font if cc > 21 else bold_font)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            target_ws.row_dimensions[experimental_row].height = 44.0
            experimental_row += 1
    signal_lines = [
        str(item or "").strip()
        for item in list(experimental_signal_audit.get("interpretation_lines") or [])
        if str(item or "").strip()
    ]
    if signal_lines:
        target_ws.merge_cells(start_row=experimental_row, start_column=21, end_row=experimental_row, end_column=24)
        signal_cell = target_ws.cell(row=experimental_row, column=21, value="Signal audit\n" + "\n".join(signal_lines))
        signal_cell.fill = copy(intro_fill)
        signal_cell.font = copy(body_font)
        signal_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        signal_cell.border = copy(thin_border)
        for cc in range(21, 25):
            target_ws.cell(row=experimental_row, column=cc).fill = copy(intro_fill)
            target_ws.cell(row=experimental_row, column=cc).border = copy(thin_border)
        target_ws.row_dimensions[experimental_row].height = 46.0
        experimental_row += 1

    system_roles_title_row = experimental_row + 1
    target_ws.merge_cells(start_row=system_roles_title_row, start_column=21, end_row=system_roles_title_row, end_column=24)
    system_roles_title = target_ws.cell(row=system_roles_title_row, column=21, value="System roles / checks")
    system_roles_title.fill = copy(section_fill)
    system_roles_title.font = copy(bold_font)
    system_roles_title.alignment = Alignment(horizontal="center", vertical="center")
    system_roles_title.border = copy(thin_border)
    for cc in range(21, 25):
        target_ws.cell(row=system_roles_title_row, column=cc).fill = copy(section_fill)
        target_ws.cell(row=system_roles_title_row, column=cc).font = copy(bold_font)
        target_ws.cell(row=system_roles_title_row, column=cc).border = copy(thin_border)
        target_ws.cell(row=system_roles_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")

    system_roles_rows = [
        ("Official row", system_official_row_role),
        ("GPRE crush proxy", system_fitted_row_role),
        ("Expanded-pass best", system_expanded_pass_role),
        ("Production winner", system_production_winner_role),
        ("Best historical fit", system_best_historical_role),
        ("Best compromise", system_best_compromise_role),
        ("Best forward lens", system_best_forward_role),
        ("Winner preview quality", system_winner_preview_quality or "n/a"),
        ("Winner forward usability", system_winner_forward_usability or "n/a"),
        ("Hedge-style study", system_hedge_role),
        ("Inconsistency detected", "Yes" if system_internal_inconsistency else "No"),
    ]
    system_roles_row = system_roles_title_row + 1
    for label_txt, value_txt in system_roles_rows:
        target_ws.merge_cells(start_row=system_roles_row, start_column=22, end_row=system_roles_row, end_column=24)
        for cc in range(21, 25):
            cell = target_ws.cell(
                row=system_roles_row,
                column=cc,
                value=label_txt if cc == 21 else value_txt if cc == 22 else None,
            )
            cell.fill = copy(zebra_fill_light if ((system_roles_row - system_roles_title_row) % 2) else zebra_fill_dark)
            cell.border = copy(thin_border)
            cell.font = copy(body_font if cc > 21 else bold_font)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        target_ws.row_dimensions[system_roles_row].height = 28.0
        system_roles_row += 1

    fp_headers = ["Quarter", "Plants", "Active regions", "Flags", "Notes", "Source refs"]
    fp_header_row = footprint_section_row + 1
    for offset, header in enumerate(fp_headers, start=2):
        cell = ws.cell(row=fp_header_row, column=offset, value=header)
        cell.fill = copy(header_fill)
        cell.font = copy(bold_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(thin_border)
    ws.row_dimensions[fp_header_row].height = 26.0
    fp_row = fp_header_row + 1
    if isinstance(footprint_df, pd.DataFrame) and not footprint_df.empty:
        for rec in footprint_df.to_dict("records"):
            quarter_val = rec.get("quarter")
            if isinstance(quarter_val, pd.Timestamp):
                quarter_val = quarter_val.date()
            vals = [
                _quarter_label_short(quarter_val) if isinstance(quarter_val, date) else str(rec.get("quarter_label") or rec.get("quarter") or ""),
                rec.get("operating_plant_count"),
                _sandbox_active_regions(rec.get("active_regions")),
                _sandbox_short_regime_flags(rec.get("regime_flags")),
                rec.get("notes"),
                rec.get("source_refs"),
            ]
            for idx, val in enumerate(vals, start=2):
                cell = ws.cell(row=fp_row, column=idx, value=val)
                cell.fill = copy(zebra_fill_light if ((fp_row - fp_header_row) % 2) else zebra_fill_dark)
                cell.font = copy(body_font)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = copy(thin_border)
            ws.row_dimensions[fp_row].height = 38.0
            fp_row += 1

    metrics_headers = ["Model", "Family", "Clean MAE", "Underlying MAE", "Hybrid", "Baseline", "Avg diff", ">2c", ">5c", "Status", "Notes"]
    metrics_header_row = fp_header_row
    for offset, header in enumerate(metrics_headers, start=9):
        cell = ws.cell(row=metrics_header_row, column=offset, value=header)
        cell.fill = copy(header_fill)
        cell.font = copy(bold_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(thin_border)
    ws.row_dimensions[metrics_header_row].height = 26.0
    metrics_row = metrics_header_row + 1
    weight_summary_map: Dict[str, str] = {}
    if isinstance(weights_df, pd.DataFrame) and not weights_df.empty:
        latest_quarter = pd.to_datetime(weights_df.get("quarter"), errors="coerce").dropna().max() if "quarter" in weights_df.columns else pd.NaT
        for model_key in sorted({str(v or "") for v in weights_df.get("model_key", []) if str(v or "").strip()}):
            sub = weights_df[weights_df["model_key"].astype(str) == model_key].copy()
            if model_key in {"optimized_weights", "plant_count_prev_quarter"}:
                sub = sub[sub["quarter"].isna()].copy()
            elif pd.notna(latest_quarter):
                sub = sub[pd.to_datetime(sub["quarter"], errors="coerce") == latest_quarter].copy()
            sub = sub.sort_values("weight", ascending=False)
            pieces = [
                f"{str(rec.get('region') or '').replace('_', ' ').title()} {float(rec.get('weight') or 0.0):.0%}"
                for rec in sub.to_dict("records")
                if float(rec.get("weight") or 0.0) > 0
            ]
            weight_summary_map[model_key] = ", ".join(pieces[:4])
    if "plant_count_weighted" in weight_summary_map:
        plant_weight_summary = weight_summary_map["plant_count_weighted"]
        for alias in (
            "bridge_current_quarter_avg",
            "bridge_front_loaded",
            "bridge_current75_prev25",
            "bridge_current50_prev50",
            "process_current_quarter_avg",
            "process_front_loaded",
            "process_current75_prev25",
            "process_current50_prev50",
            "hedge_disclosed_bridge_prior_current",
            "hedge_disclosed_bridge_prior_front",
            "hedge_disclosed_process_prior_current",
            "hedge_disclosed_process_prior_front",
            "hedge_pattern_bridge_prior_current",
            "hedge_pattern_bridge_prior_front",
            "hedge_pattern_process_prior_current",
            "hedge_pattern_process_prior_front",
            "bid_adjusted_offset",
        ):
            weight_summary_map.setdefault(alias, plant_weight_summary)
    weight_summary_map.setdefault("simple_market", "Delivered corn row; active-capacity weighted basis plus gas")
    if isinstance(leaderboard_df, pd.DataFrame) and not leaderboard_df.empty:
        metrics_view = leaderboard_df.copy().sort_values(["chosen", "hybrid_score", "underlying_mae", "clean_mae"], ascending=[False, True, True, True], na_position="last")
        for rec in metrics_view.to_dict("records"):
            status_bits = []
            if bool(rec.get("production_winner")) or bool(rec.get("chosen")):
                status_bits.append("Winner")
            if bool(rec.get("expanded_best_candidate")):
                status_bits.append("Expanded best")
            chosen_flag = " + ".join(status_bits)
            baseline_status = str(rec.get("baseline_status") or "").strip()
            baseline_display = {
                "incumbent_current_state": "Incumbent",
                "incumbent_process_comparator": "Process comp",
                "new_candidate": "New",
                "existing_candidate": "Existing",
            }.get(baseline_status, baseline_status.replace("_", " ").title())
            note_bits = []
            if bool(rec.get("comparison_only")):
                note_bits.append("Comparison only")
            guard_reason = str(rec.get("selection_guard_reason") or "").strip()
            if guard_reason:
                note_bits.append(_sandbox_guard_label(guard_reason))
            promotion_reason = str(rec.get("promotion_guard_reason") or "").strip()
            if promotion_reason:
                note_bits.append("Promotion: " + _sandbox_guard_label(promotion_reason))
            incremental_value_status = str(rec.get("incremental_value_status") or "").strip()
            if incremental_value_status:
                note_bits.append(f"Incremental value {incremental_value_status}")
            q_mae_bits = []
            for quarter_label, quarter_val in (
                ("Q1", rec.get("q1_mae")),
                ("Q2", rec.get("q2_mae")),
                ("Q3", rec.get("q3_mae")),
                ("Q4", rec.get("q4_mae")),
            ):
                quarter_num = pd.to_numeric(quarter_val, errors="coerce")
                if pd.notna(quarter_num):
                    q_mae_bits.append(f"{quarter_label} {float(quarter_num):.3f}")
            if q_mae_bits:
                note_bits.append("Quarter MAE: " + ", ".join(q_mae_bits))
            bias_num = pd.to_numeric(rec.get("q1_mean_error"), errors="coerce")
            if pd.notna(bias_num):
                note_bits.append(f"Q1 bias {float(bias_num):.3f}")
            bias_direction = str(rec.get("bias_direction") or "").strip()
            if bias_direction:
                note_bits.append(f"Bias {bias_direction}")
            corr_num = pd.to_numeric(rec.get("test_corr"), errors="coerce")
            if pd.notna(corr_num):
                note_bits.append(f"Test corr {float(corr_num):.3f}")
            sign_num = pd.to_numeric(rec.get("test_sign_hit_rate"), errors="coerce")
            if pd.notna(sign_num):
                note_bits.append(f"Sign {float(sign_num):.0%}")
            preview_quality_status = str(rec.get("live_preview_quality_status") or "").strip()
            if preview_quality_status:
                preview_mae_num = pd.to_numeric(rec.get("live_preview_mae"), errors="coerce")
                preview_max_num = pd.to_numeric(rec.get("live_preview_max_error"), errors="coerce")
                preview_worst_phase = str(rec.get("live_preview_worst_phase") or "").strip()
                preview_txt = f"Preview {preview_quality_status}"
                if pd.notna(preview_mae_num):
                    preview_txt += f" | MAE {float(preview_mae_num):.3f}"
                if pd.notna(preview_max_num):
                    preview_txt += f" | Max {float(preview_max_num):.3f}"
                if preview_worst_phase:
                    preview_txt += f" | Main mode {_sandbox_preview_phase_label(preview_worst_phase)}"
                note_bits.append(preview_txt)
            promotion_failures = str(rec.get("promotion_guard_failures") or "").strip()
            if "live_preview_quality_not_faithful_enough" in promotion_failures:
                note_bits.append("Preview block: not faithful enough to the full model")
            hard_mae_num = pd.to_numeric(rec.get("hard_quarter_mae"), errors="coerce")
            hard_count_num = pd.to_numeric(rec.get("hard_quarter_count"), errors="coerce")
            if pd.notna(hard_mae_num):
                hard_txt = f"Hard-quarter MAE {float(hard_mae_num):.3f}"
                if pd.notna(hard_count_num):
                    hard_txt += f" ({int(hard_count_num)} qtrs)"
                note_bits.append(hard_txt)
            top_miss = str(rec.get("top_miss_quarters") or "").strip()
            if top_miss:
                note_bits.append(f"Top misses: {top_miss}")
            preview_top_miss = str(rec.get("live_preview_top_miss_quarters") or "").strip()
            if preview_top_miss:
                note_bits.append(f"Preview misses: {preview_top_miss}")
            hard_top_miss = str(rec.get("hard_quarter_top_miss_quarters") or "").strip()
            if hard_top_miss:
                note_bits.append(f"Hard-quarter misses: {hard_top_miss}")
            weight_note = weight_summary_map.get(str(rec.get("model_key") or ""), "")
            if weight_note:
                note_bits.append(weight_note)
            vals = [
                _sandbox_model_label(rec.get("model_key")),
                str(rec.get("family_label") or rec.get("family") or ""),
                rec.get("clean_mae"),
                rec.get("underlying_mae"),
                rec.get("hybrid_score"),
                baseline_display,
                rec.get("avg_abs_diff_vs_official"),
                rec.get("diff_quarters_gt_0_02_vs_official"),
                rec.get("diff_quarters_gt_0_05_vs_official"),
                chosen_flag,
                " | ".join(note_bits),
            ]
            for idx, val in enumerate(vals, start=9):
                cell = ws.cell(row=metrics_row, column=idx, value=val)
                cell.fill = copy(zebra_fill_light if ((metrics_row - metrics_header_row) % 2) else zebra_fill_dark)
                cell.font = copy(body_font)
                cell.alignment = Alignment(horizontal="left" if idx in {9, 10, 19} else "center", vertical="top", wrap_text=idx in {9, 10, 19})
                cell.border = copy(thin_border)
                if idx in {11, 12, 13, 15} and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0.000"
            ws.row_dimensions[metrics_row].height = 36.0
            metrics_row += 1

    recent_compare_title_row = metrics_row + 1
    ws.merge_cells(start_row=recent_compare_title_row, start_column=9, end_row=recent_compare_title_row, end_column=17)
    recent_compare_title = ws.cell(row=recent_compare_title_row, column=9, value="Recent-quarter winner comparison")
    recent_compare_title.fill = copy(section_fill)
    recent_compare_title.font = copy(bold_font)
    recent_compare_title.alignment = Alignment(horizontal="center", vertical="center")
    recent_compare_title.border = copy(thin_border)
    for cc in range(9, 18):
        ws.cell(row=recent_compare_title_row, column=cc).fill = copy(section_fill)
        ws.cell(row=recent_compare_title_row, column=cc).font = copy(bold_font)
        ws.cell(row=recent_compare_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=recent_compare_title_row, column=cc).border = copy(thin_border)
    recent_compare_header_row = recent_compare_title_row + 1
    recent_headers = ["Quarter", "Official", "Incumbent", "Process comp", "Expanded best", "Winner", "Target", "Winner err", "Hard?"]
    for offset_idx, header in enumerate(recent_headers, start=9):
        cell = ws.cell(row=recent_compare_header_row, column=offset_idx, value=header)
        cell.fill = copy(header_fill)
        cell.font = copy(bold_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(thin_border)
    ws.row_dimensions[recent_compare_header_row].height = 28.0
    recent_compare_row = recent_compare_header_row + 1
    if isinstance(recent_quarter_comparison_df, pd.DataFrame) and not recent_quarter_comparison_df.empty:
        for rec in recent_quarter_comparison_df.to_dict("records"):
            vals = [
                rec.get("Quarter"),
                rec.get("Official"),
                rec.get("Incumbent"),
                rec.get("Process comp"),
                rec.get("Expanded best"),
                rec.get("Winner"),
                rec.get("Target"),
                rec.get("Winner err"),
                rec.get("Hard?"),
            ]
            for idx, val in enumerate(vals, start=9):
                cell = ws.cell(row=recent_compare_row, column=idx, value=val)
                cell.fill = copy(zebra_fill_light if ((recent_compare_row - recent_compare_header_row) % 2) else zebra_fill_dark)
                cell.font = copy(body_font)
                cell.alignment = Alignment(horizontal="left" if idx in {9, 17} else "center", vertical="top", wrap_text=idx in {9, 17})
                cell.border = copy(thin_border)
                if idx in {10, 11, 12, 13, 14, 15, 16} and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0.000"
            ws.row_dimensions[recent_compare_row].height = 30.0
            recent_compare_row += 1
    else:
        ws.merge_cells(start_row=recent_compare_row, start_column=9, end_row=recent_compare_row, end_column=17)
        recent_empty = ws.cell(row=recent_compare_row, column=9, value="No recent-quarter comparison rows were available.")
        recent_empty.fill = copy(intro_fill)
        recent_empty.font = copy(body_font)
        recent_empty.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        recent_empty.border = copy(thin_border)
        for cc in range(9, 18):
            ws.cell(row=recent_compare_row, column=cc).fill = copy(intro_fill)
            ws.cell(row=recent_compare_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[recent_compare_row].height = 26.0
        recent_compare_row += 1

    offset_section_row = max(fp_row, recent_compare_row) + 1
    ws.merge_cells(start_row=offset_section_row, start_column=2, end_row=offset_section_row, end_column=15)
    offset_title = ws.cell(row=offset_section_row, column=2, value="Current GPRE bids vs AMS reference offsets")
    offset_title.fill = copy(section_fill)
    offset_title.font = copy(bold_font)
    offset_title.alignment = Alignment(horizontal="center", vertical="center")
    offset_title.border = copy(thin_border)
    for cc in range(2, 16):
        ws.cell(row=offset_section_row, column=cc).fill = copy(section_fill)
        ws.cell(row=offset_section_row, column=cc).font = copy(bold_font)
        ws.cell(row=offset_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=offset_section_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[offset_section_row].height = 22.0

    offset_header_row = offset_section_row + 1
    offset_headers = [
        "Region",
        "GPRE bid c/bu",
        "AMS ref c/bu",
        "Offset c/bu",
        "Locations",
        "AMS reference",
    ]
    for offset_idx, header in enumerate(offset_headers, start=2):
        cell = ws.cell(row=offset_header_row, column=offset_idx, value=header)
        cell.fill = copy(header_fill)
        cell.font = copy(bold_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(thin_border)
    ws.row_dimensions[offset_header_row].height = 28.0

    offset_data_row = offset_header_row + 1
    if isinstance(bid_adjusted_offsets_df, pd.DataFrame) and not bid_adjusted_offsets_df.empty:
        for rec in bid_adjusted_offsets_df.to_dict("records"):
            vals = [
                str(rec.get("region") or "").replace("_", " ").title(),
                rec.get("gpre_bid_basis_cents_per_bu"),
                rec.get("ams_reference_basis_cents_per_bu"),
                rec.get("offset_cents_per_bu"),
                rec.get("locations"),
                (
                    f"{str(rec.get('reference_method') or '').strip()}"
                    if not str(rec.get("reference_as_of") or "").strip()
                    else f"{str(rec.get('reference_method') or '').strip()} | As of {_overlay_market_date_text(rec.get('reference_as_of'))}"
                ),
            ]
            for idx, val in enumerate(vals, start=2):
                cell = ws.cell(row=offset_data_row, column=idx, value=val)
                cell.fill = copy(zebra_fill_light if ((offset_data_row - offset_header_row) % 2) else zebra_fill_dark)
                cell.font = copy(body_font)
                cell.alignment = Alignment(horizontal="left" if idx in {2, 6, 7} else "center", vertical="top", wrap_text=idx in {6, 7})
                cell.border = copy(thin_border)
                if idx in {3, 4, 5} and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0.0"
            ws.row_dimensions[offset_data_row].height = 32.0
            offset_data_row += 1
    else:
        ws.merge_cells(start_row=offset_data_row, start_column=2, end_row=offset_data_row, end_column=15)
        unavailable_text = (
            "No local or live GPRE bid snapshot was available for the bid-adjusted offset comparison."
            if not isinstance(gpre_bid_snapshot, dict) or str(gpre_bid_snapshot.get("status") or "").strip().lower() != "ok"
            else "GPRE bid snapshot was available, but no AMS regional reference could be constructed for the offset comparison."
        )
        cell = ws.cell(row=offset_data_row, column=2, value=unavailable_text)
        cell.fill = copy(intro_fill)
        cell.font = copy(body_font)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = copy(thin_border)
        for cc in range(2, 16):
            ws.cell(row=offset_data_row, column=cc).fill = copy(intro_fill)
            ws.cell(row=offset_data_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[offset_data_row].height = 30.0
        offset_data_row += 1

    table_start_row = offset_data_row + 1
    ws.merge_cells(start_row=table_start_row, start_column=2, end_row=table_start_row, end_column=20)
    table_title = ws.cell(row=table_start_row, column=2, value="Quarterly comparison table")
    table_title.fill = copy(section_fill)
    table_title.font = copy(bold_font)
    table_title.alignment = Alignment(horizontal="center", vertical="center")
    table_title.border = copy(thin_border)
    for cc in range(2, 21):
        ws.cell(row=table_start_row, column=cc).fill = copy(section_fill)
        ws.cell(row=table_start_row, column=cc).font = copy(bold_font)
        ws.cell(row=table_start_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=table_start_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[table_start_row].height = 22.0

    quarterly_headers = [
        "Quarter",
        "Split",
        "Target type",
        "Regime flags",
        "Target $/gal",
        "Simple market",
        "GPRE proxy",
        "Underlying",
        "Reported",
        "Simple err",
        "GPRE err",
        "Coverage notes",
        "Denominator",
        "Chosen model",
    ]
    quarterly_cols = [
        "quarter_label",
        "train_test_flag",
        "target_basis",
        "regime_flags",
        "evaluation_target_margin_usd_per_gal",
        "simple_market_proxy_usd_per_gal",
        "gpre_proxy_official_usd_per_gal",
        "underlying_crush_margin_usd_per_gal",
        "reported_consolidated_crush_margin_usd_per_gal",
        "simple_market_proxy_error_usd_per_gal",
        "gpre_proxy_error_usd_per_gal",
        "coverage_notes",
        "denominator_policy",
        "gpre_proxy_model_key",
    ]
    quarterly_header_row = table_start_row + 1
    for offset, header in enumerate(quarterly_headers, start=2):
        cell = ws.cell(row=quarterly_header_row, column=offset, value=header)
        cell.fill = copy(header_fill)
        cell.font = copy(bold_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(thin_border)
    ws.row_dimensions[quarterly_header_row].height = 34.0

    data_row = quarterly_header_row + 1
    if isinstance(quarterly_df, pd.DataFrame) and not quarterly_df.empty:
        for rec in quarterly_df.to_dict("records"):
            for offset, col_name in enumerate(quarterly_cols, start=2):
                val = rec.get(col_name)
                if col_name == "target_basis":
                    val = str(val or "").strip().title()
                if col_name == "regime_flags":
                    val = _sandbox_short_regime_flags(val)
                elif col_name == "coverage_notes":
                    val = _sandbox_short_coverage(val)
                elif col_name == "denominator_policy":
                    val = _sandbox_short_denominator(val)
                elif col_name == "gpre_proxy_model_key":
                    val = _sandbox_model_label(val)
                cell = ws.cell(row=data_row, column=offset, value=val)
                cell.fill = copy(zebra_fill_light if ((data_row - quarterly_header_row) % 2) else zebra_fill_dark)
                cell.font = copy(body_font)
                cell.alignment = Alignment(horizontal="left" if offset in {2, 3, 4, 12, 13, 14} else "center", vertical="top", wrap_text=offset in {4, 12, 13, 14})
                cell.border = copy(thin_border)
                if offset in {6, 7, 8, 9, 10, 11} and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0.000"
            ws.row_dimensions[data_row].height = 32.0
            data_row += 1

    process_build_section_row = data_row + 1
    approx_market_crush_build_up_layout = _write_gpre_approx_market_crush_build_up_section(
        ws,
        process_build_section_row,
        source_sheet_name="Economics_Overlay",
    )
    coproduct_section_row = int(approx_market_crush_build_up_layout.get("next_row") or process_build_section_row) + 2
    coproduct_history_layout: Dict[str, Any] = {}
    if overlay_source_ws is not None:
        distillers_price_state = _overlay_row_readiness_state(market_rows.get("distillers_grains_price"))
        direct_corn_oil_price_state = _overlay_row_readiness_state(market_rows.get("renewable_corn_oil_price"))
        resolved_corn_oil_price_state = dict(direct_corn_oil_price_state)
        distillers_yield_state = _overlay_constant_input_state(coeff_rows.get("distillers_yield"))
        uhp_yield_state = _overlay_constant_input_state(coeff_rows.get("uhp_yield"))
        corn_oil_yield_state = _overlay_constant_input_state(coeff_rows.get("renewable_corn_oil_yield"))
        distillers_contribution_state = _require_coproduct_state(distillers_yield_state, distillers_price_state)
        uhp_contribution_state = _require_coproduct_state(uhp_yield_state, _overlay_row_readiness_state(market_rows.get("uhp_price")))
        corn_oil_contribution_state = _require_coproduct_state(corn_oil_yield_state, resolved_corn_oil_price_state)
        approximate_coproduct_credit_state = _combine_coproduct_state(
            distillers_contribution_state,
            uhp_contribution_state,
            corn_oil_contribution_state,
        )
        nwer_coproduct_source_state = _coproduct_source_state(
            source_type_prefix="nwer",
            series_prefixes=("corn_oil_", "ddgs_10_"),
        )
        ams_3618_coproduct_source_state = _coproduct_source_state(
            source_type_prefix="ams_3618",
            series_prefixes=("corn_oil_", "ddgs_10_"),
        )
        corn_oil_source_mode_text = _overlay_source_mode_text(market_rows.get("renewable_corn_oil_price"))
        distillers_source_mode_text = _overlay_source_mode_text(market_rows.get("distillers_grains_price"))
        current_resolved_workbook_source = _classify_coproduct_resolved_source(
            corn_oil_source_mode_text,
            distillers_source_mode_text,
        )
        coproduct_history_records = _quarterly_coproduct_history_records()
        coproduct_frame_summary_records = _coproduct_frame_summary_records(coproduct_history_records)
        current_frame_summary_record = next(
            (
                dict(rec)
                for rec in list(coproduct_frame_summary_records or [])
                if str((rec or {}).get("frame_key") or "").strip() == "current_qtd"
            ),
            {},
        )
        current_resolved_workbook_source = str(
            current_frame_summary_record.get("resolved_source_mode") or current_resolved_workbook_source or "Unknown/blank"
        )
        nwer_coproduct_gate_pass = any(bool(nwer_coproduct_source_state.get(bucket)) for bucket in ("historical", "current", "next"))
        ams_3618_coproduct_gate_pass = any(bool(ams_3618_coproduct_source_state.get(bucket)) for bucket in ("historical", "current", "next"))
        resolved_corn_oil_price_gate_pass = any(bool(resolved_corn_oil_price_state.get(bucket)) for bucket in ("historical", "current", "next"))
        distillers_price_gate_pass = any(bool(distillers_price_state.get(bucket)) for bucket in ("historical", "current", "next"))
        approximate_coproduct_credit_gate_pass = any(
            bool(approximate_coproduct_credit_state.get(bucket)) for bucket in ("historical", "current", "next")
        )
        overlay_activation_gate_pass = bool(
            nwer_coproduct_gate_pass
            and bool(resolved_corn_oil_price_state.get("current"))
            and bool(distillers_price_state.get("current"))
            and bool(approximate_coproduct_credit_state.get("current"))
        )
        corn_oil_gate_specs = [
            (
                "NWER coproduct rows",
                "YES" if nwer_coproduct_gate_pass else "NO",
                (
                    "Parsed NWER export rows now include direct corn-oil or DDGS coproduct series."
                    if nwer_coproduct_gate_pass
                    else "No parsed NWER coproduct rows were found in the current GPRE export."
                ),
            ),
            (
                "AMS 3618 coproduct rows",
                "YES" if ams_3618_coproduct_gate_pass else "NO",
                (
                    "Parsed AMS 3618 export rows now include direct corn-oil or DDGS coproduct series. Secondary/corroborating manual fallback/backfill source; not required for visible activation."
                    if ams_3618_coproduct_gate_pass
                    else "No parsed AMS 3618 coproduct rows were found in the current GPRE export. Secondary/corroborating manual fallback/backfill only; not required for the first NWER-backed visible block."
                ),
            ),
            (
                "Renewable corn oil price",
                "YES" if resolved_corn_oil_price_gate_pass else "NO",
                (
                    "Resolved non-blank from direct parsed market rows."
                    if resolved_corn_oil_price_gate_pass
                    else "Direct market row is still blank in the live GPRE path."
                ),
            ),
            (
                "Distillers grains price",
                "YES" if distillers_price_gate_pass else "NO",
                (
                    "Resolved non-blank from direct parsed market rows."
                    if distillers_price_gate_pass
                    else "Direct DDGS market row is still blank in the live GPRE path."
                ),
            ),
            (
                "Approximate coproduct credit",
                "YES" if approximate_coproduct_credit_gate_pass else "NO",
                (
                    "Displayed sandbox contribution cells resolve non-blank."
                    if approximate_coproduct_credit_gate_pass
                    else "Displayed sandbox contribution cells are still blocked in the live GPRE path."
                ),
            ),
            (
                "Overlay activation",
                "GO" if overlay_activation_gate_pass else "HOLD",
                (
                    "NWER current-quarter sufficiency is enough for visible activation; AMS 3618 remains secondary and may still be the current resolved workbook source."
                    if overlay_activation_gate_pass
                    else "Keep the visible block deferred until NWER-backed current corn-oil, DDGS, and coproduct-credit cells are all non-blank."
                ),
            ),
        ]
        coproduct_provenance_specs = [
            (
                "Primary live activation source",
                "NWER",
                "Current-quarter NWER sufficiency controls GO/HOLD for the first visible block.",
            ),
            (
                "Secondary corroborating source",
                "AMS 3618",
                "Manual fallback/backfill and corroboration; not required for visible activation.",
            ),
            (
                "Current resolved workbook source",
                current_resolved_workbook_source,
                "Shows which source currently feeds the visible price rows.",
            ),
        ]
        corn_oil_gate_section_row = coproduct_section_row
        ws.merge_cells(start_row=corn_oil_gate_section_row, start_column=2, end_row=corn_oil_gate_section_row, end_column=15)
        corn_oil_gate_title = ws.cell(row=corn_oil_gate_section_row, column=2, value="Coproduct source gate")
        corn_oil_gate_title.fill = copy(section_fill)
        corn_oil_gate_title.font = copy(bold_font)
        corn_oil_gate_title.alignment = Alignment(horizontal="center", vertical="center")
        corn_oil_gate_title.border = copy(thin_border)
        for cc in range(2, 16):
            ws.cell(row=corn_oil_gate_section_row, column=cc).fill = copy(section_fill)
            ws.cell(row=corn_oil_gate_section_row, column=cc).font = copy(bold_font)
            ws.cell(row=corn_oil_gate_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=corn_oil_gate_section_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[corn_oil_gate_section_row].height = 22.0

        corn_oil_gate_header_row = corn_oil_gate_section_row + 1
        corn_oil_gate_spans = [
            (2, 5, "Gate"),
            (6, 7, "Status"),
            (8, 15, "Reason"),
        ]
        for start_col, end_col, header_txt in corn_oil_gate_spans:
            if end_col > start_col:
                ws.merge_cells(start_row=corn_oil_gate_header_row, start_column=start_col, end_row=corn_oil_gate_header_row, end_column=end_col)
            for cc in range(start_col, end_col + 1):
                ws.cell(row=corn_oil_gate_header_row, column=cc).fill = copy(header_fill)
                ws.cell(row=corn_oil_gate_header_row, column=cc).font = copy(bold_font)
                ws.cell(row=corn_oil_gate_header_row, column=cc).border = copy(thin_border)
                ws.cell(row=corn_oil_gate_header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=corn_oil_gate_header_row, column=start_col, value=header_txt)
        ws.row_dimensions[corn_oil_gate_header_row].height = 24.0

        corn_oil_gate_row = corn_oil_gate_header_row + 1
        corn_oil_gate_rows: Dict[str, int] = {}
        for gate_label, gate_status, gate_reason in corn_oil_gate_specs:
            row_fill = copy(zebra_fill_light if ((corn_oil_gate_row - corn_oil_gate_header_row) % 2) else zebra_fill_dark)
            gate_spans_and_values = [
                (2, 5, gate_label),
                (6, 7, gate_status),
                (8, 15, gate_reason),
            ]
            for start_col, end_col, value_txt in gate_spans_and_values:
                if end_col > start_col:
                    ws.merge_cells(start_row=corn_oil_gate_row, start_column=start_col, end_row=corn_oil_gate_row, end_column=end_col)
                cell = ws.cell(row=corn_oil_gate_row, column=start_col, value=value_txt)
                cell.fill = copy(row_fill)
                cell.font = copy(body_font)
                cell.border = copy(thin_border)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                for cc in range(start_col, end_col + 1):
                    ws.cell(row=corn_oil_gate_row, column=cc).fill = copy(row_fill)
                    ws.cell(row=corn_oil_gate_row, column=cc).border = copy(thin_border)
                    ws.cell(row=corn_oil_gate_row, column=cc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[corn_oil_gate_row].height = 28.0
            corn_oil_gate_rows[str(gate_label or "")] = corn_oil_gate_row
            corn_oil_gate_row += 1

        corn_oil_gate_note_row = corn_oil_gate_row
        ws.merge_cells(start_row=corn_oil_gate_note_row, start_column=2, end_row=corn_oil_gate_note_row, end_column=15)
        corn_oil_gate_note = ws.cell(
            row=corn_oil_gate_note_row,
            column=2,
            value=(
                "Stage B.4 keeps NWER as the primary live activation source for the first visible coproduct block. "
                "AMS 3618 remains secondary/corroborating and is best used as a manual fallback/backfill source. "
                "Current resolved workbook price rows may come from either source without changing the activation rule. "
                "AMS 3511 remains deferred/manual."
            ),
        )
        corn_oil_gate_note.fill = copy(intro_fill)
        corn_oil_gate_note.font = copy(body_font)
        corn_oil_gate_note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        corn_oil_gate_note.border = copy(thin_border)
        for cc in range(2, 16):
            ws.cell(row=corn_oil_gate_note_row, column=cc).fill = copy(intro_fill)
            ws.cell(row=corn_oil_gate_note_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[corn_oil_gate_note_row].height = 34.0
        corn_oil_gate_check_layout = {
            "title_row": corn_oil_gate_section_row,
            "header_row": corn_oil_gate_header_row,
            "gate_rows": corn_oil_gate_rows,
            "note_row": corn_oil_gate_note_row,
            "next_row": corn_oil_gate_note_row,
            "overlay_activation_pass": overlay_activation_gate_pass,
        }
        coproduct_provenance_section_row = corn_oil_gate_note_row + 1
        ws.merge_cells(start_row=coproduct_provenance_section_row, start_column=2, end_row=coproduct_provenance_section_row, end_column=15)
        provenance_title = ws.cell(row=coproduct_provenance_section_row, column=2, value="Source provenance")
        provenance_title.fill = copy(section_fill)
        provenance_title.font = copy(bold_font)
        provenance_title.alignment = Alignment(horizontal="center", vertical="center")
        provenance_title.border = copy(thin_border)
        for cc in range(2, 16):
            ws.cell(row=coproduct_provenance_section_row, column=cc).fill = copy(section_fill)
            ws.cell(row=coproduct_provenance_section_row, column=cc).font = copy(bold_font)
            ws.cell(row=coproduct_provenance_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=coproduct_provenance_section_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[coproduct_provenance_section_row].height = 22.0

        coproduct_provenance_header_row = coproduct_provenance_section_row + 1
        provenance_spans = [
            (2, 5, "Role"),
            (6, 7, "Source"),
            (8, 15, "Meaning"),
        ]
        for start_col, end_col, header_txt in provenance_spans:
            if end_col > start_col:
                ws.merge_cells(start_row=coproduct_provenance_header_row, start_column=start_col, end_row=coproduct_provenance_header_row, end_column=end_col)
            for cc in range(start_col, end_col + 1):
                ws.cell(row=coproduct_provenance_header_row, column=cc).fill = copy(header_fill)
                ws.cell(row=coproduct_provenance_header_row, column=cc).font = copy(bold_font)
                ws.cell(row=coproduct_provenance_header_row, column=cc).border = copy(thin_border)
                ws.cell(row=coproduct_provenance_header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=coproduct_provenance_header_row, column=start_col, value=header_txt)
        ws.row_dimensions[coproduct_provenance_header_row].height = 24.0

        coproduct_provenance_row = coproduct_provenance_header_row + 1
        provenance_rows: Dict[str, int] = {}
        for role_label, source_label, source_reason in coproduct_provenance_specs:
            row_fill = copy(zebra_fill_light if ((coproduct_provenance_row - coproduct_provenance_header_row) % 2) else zebra_fill_dark)
            provenance_spans_and_values = [
                (2, 5, role_label),
                (6, 7, source_label),
                (8, 15, source_reason),
            ]
            for start_col, end_col, value_txt in provenance_spans_and_values:
                if end_col > start_col:
                    ws.merge_cells(start_row=coproduct_provenance_row, start_column=start_col, end_row=coproduct_provenance_row, end_column=end_col)
                cell = ws.cell(row=coproduct_provenance_row, column=start_col, value=value_txt)
                cell.fill = copy(row_fill)
                cell.font = copy(body_font)
                cell.border = copy(thin_border)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                for cc in range(start_col, end_col + 1):
                    ws.cell(row=coproduct_provenance_row, column=cc).fill = copy(row_fill)
                    ws.cell(row=coproduct_provenance_row, column=cc).border = copy(thin_border)
                    ws.cell(row=coproduct_provenance_row, column=cc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[coproduct_provenance_row].height = 28.0
            provenance_rows[str(role_label or "")] = coproduct_provenance_row
            coproduct_provenance_row += 1
        corn_oil_gate_check_layout["provenance_rows"] = provenance_rows
        corn_oil_gate_check_layout["provenance_title_row"] = coproduct_provenance_section_row
        corn_oil_gate_check_layout["provenance_header_row"] = coproduct_provenance_header_row
        corn_oil_gate_check_layout["current_resolved_workbook_source"] = current_resolved_workbook_source
        corn_oil_gate_check_layout["next_row"] = coproduct_provenance_row - 1
        coproduct_section_row = coproduct_provenance_row + 1
        coproduct_specs = [
            (
                "Renewable corn oil price",
                "Direct market",
                "Direct from parsed rows",
                resolved_corn_oil_price_state,
                "market",
            ),
            (
                "Distillers grains price",
                "Direct market",
                "Direct from parsed rows",
                distillers_price_state,
                "market",
            ),
            (
                "NWER coproduct rows",
                "Weekly bioenergy",
                "Provider source",
                nwer_coproduct_source_state,
                "market",
            ),
            (
                "AMS 3618 coproduct rows",
                "Weekly co-products",
                "Provider source",
                ams_3618_coproduct_source_state,
                "market",
            ),
            (
                "Approximate coproduct credit",
                "Derived build-up",
                "Contribution sum",
                approximate_coproduct_credit_state,
                "derived",
            ),
        ]
        ws.merge_cells(start_row=coproduct_section_row, start_column=2, end_row=coproduct_section_row, end_column=15)
        coproduct_title = ws.cell(row=coproduct_section_row, column=2, value="Coproduct signal readiness")
        coproduct_title.fill = copy(section_fill)
        coproduct_title.font = copy(bold_font)
        coproduct_title.alignment = Alignment(horizontal="center", vertical="center")
        coproduct_title.border = copy(thin_border)
        for cc in range(2, 16):
            ws.cell(row=coproduct_section_row, column=cc).fill = copy(section_fill)
            ws.cell(row=coproduct_section_row, column=cc).font = copy(bold_font)
            ws.cell(row=coproduct_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=coproduct_section_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[coproduct_section_row].height = 22.0

        coproduct_note_row = coproduct_section_row + 1
        ws.merge_cells(start_row=coproduct_note_row, start_column=2, end_row=coproduct_note_row, end_column=15)
        coproduct_note = ws.cell(
            row=coproduct_note_row,
            column=2,
            value=(
                "Stage B.4 keeps NWER as the sufficient first visible coproduct source. "
                "AMS 3618 remains secondary/corroborating/manual fallback-backfill, current resolved price rows may come from either source, and 3511 remains deferred/manual."
            ),
        )
        coproduct_note.fill = copy(intro_fill)
        coproduct_note.font = copy(body_font)
        coproduct_note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        coproduct_note.border = copy(thin_border)
        for cc in range(2, 16):
            ws.cell(row=coproduct_note_row, column=cc).fill = copy(intro_fill)
            ws.cell(row=coproduct_note_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[coproduct_note_row].height = 34.0

        coproduct_header_row = coproduct_note_row + 1
        coproduct_header_spans = [
            (2, 3, "Signal"),
            (4, 5, "Source mode"),
            (6, 7, "Status"),
            (8, 9, "Filled now?"),
            (10, 11, "Historical"),
            (12, 13, "Current"),
            (14, 15, "Next"),
        ]
        for start_col, end_col, header_txt in coproduct_header_spans:
            if end_col > start_col:
                ws.merge_cells(start_row=coproduct_header_row, start_column=start_col, end_row=coproduct_header_row, end_column=end_col)
            for cc in range(start_col, end_col + 1):
                ws.cell(row=coproduct_header_row, column=cc).fill = copy(header_fill)
                ws.cell(row=coproduct_header_row, column=cc).font = copy(bold_font)
                ws.cell(row=coproduct_header_row, column=cc).border = copy(thin_border)
                ws.cell(row=coproduct_header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=coproduct_header_row, column=start_col, value=header_txt)
        ws.row_dimensions[coproduct_header_row].height = 26.0

        coproduct_row = coproduct_header_row + 1
        coproduct_signal_rows: Dict[str, int] = {}
        for signal_label, source_mode_txt, status_txt, state_map, readiness_kind in coproduct_specs:
            row_fill = copy(zebra_fill_light if ((coproduct_row - coproduct_header_row) % 2) else zebra_fill_dark)
            spans_and_values = [
                (2, 3, signal_label),
                (4, 5, source_mode_txt),
                (6, 7, status_txt),
                (8, 9, _coproduct_filled_now_text(state_map)),
                (10, 11, _coproduct_readiness_bucket_text(state_map, "historical", readiness_kind)),
                (12, 13, _coproduct_readiness_bucket_text(state_map, "current", readiness_kind)),
                (14, 15, _coproduct_readiness_bucket_text(state_map, "next", readiness_kind)),
            ]
            for start_col, end_col, value_txt in spans_and_values:
                if end_col > start_col:
                    ws.merge_cells(start_row=coproduct_row, start_column=start_col, end_row=coproduct_row, end_column=end_col)
                cell = ws.cell(row=coproduct_row, column=start_col, value=value_txt)
                cell.fill = copy(row_fill)
                cell.font = copy(body_font)
                cell.border = copy(thin_border)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                for cc in range(start_col, end_col + 1):
                    ws.cell(row=coproduct_row, column=cc).fill = copy(row_fill)
                    ws.cell(row=coproduct_row, column=cc).border = copy(thin_border)
                    ws.cell(row=coproduct_row, column=cc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[coproduct_row].height = 28.0
            coproduct_signal_rows[str(signal_label or "")] = coproduct_row
            coproduct_row += 1
        coproduct_signal_readiness_layout = {
            "title_row": coproduct_section_row,
            "note_row": coproduct_note_row,
            "header_row": coproduct_header_row,
            "signal_rows": coproduct_signal_rows,
            "next_row": coproduct_row - 1,
        }
        if coproduct_frame_summary_records:
            frame_section_row = int(coproduct_signal_readiness_layout.get("next_row") or (coproduct_row - 1)) + 2
            frame_start_col = 2  # B
            frame_end_col = 10  # J
            frame_note_row = frame_section_row + 1
            frame_header_row = frame_section_row + 2
            ws.merge_cells(start_row=frame_section_row, start_column=frame_start_col, end_row=frame_section_row, end_column=frame_end_col)
            frame_title = ws.cell(row=frame_section_row, column=frame_start_col, value="Coproduct frame summary")
            frame_title.fill = copy(section_fill)
            frame_title.font = copy(bold_font)
            frame_title.alignment = Alignment(horizontal="center", vertical="center")
            frame_title.border = copy(thin_border)
            for cc in range(frame_start_col, frame_end_col + 1):
                ws.cell(row=frame_section_row, column=cc).fill = copy(section_fill)
                ws.cell(row=frame_section_row, column=cc).font = copy(bold_font)
                ws.cell(row=frame_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=frame_section_row, column=cc).border = copy(thin_border)
            ws.row_dimensions[frame_section_row].height = 22.0

            ws.merge_cells(start_row=frame_note_row, start_column=frame_start_col, end_row=frame_note_row, end_column=frame_end_col)
            frame_note = ws.cell(
                row=frame_note_row,
                column=frame_start_col,
                value=(
                    "Frame values use quarter-aware active-capacity weighting across the active GPRE footprint. "
                    "DDGS uses all active plants; corn oil uses the same active-footprint weighting as a producer-subset approximation because plant-level corn-oil flags are not available. "
                    "Quarter-open uses early-quarter observations when present and otherwise carries prior-quarter values; next quarter outlook freezes the resolved quarter-open weighted value and only falls back to prior quarter when quarter-open is unavailable. "
                    "Coverage is covered active-capacity share; frame $m uses implied company-gallons scaling, with latest reported coproduct/corn volumes used as the forward yield anchor when available."
                ),
            )
            frame_note.fill = copy(intro_fill)
            frame_note.font = copy(body_font)
            frame_note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            frame_note.border = copy(thin_border)
            for cc in range(frame_start_col, frame_end_col + 1):
                ws.cell(row=frame_note_row, column=cc).fill = copy(intro_fill)
                ws.cell(row=frame_note_row, column=cc).border = copy(thin_border)
            ws.row_dimensions[frame_note_row].height = 58.0

            frame_headers = {
                2: "Frame",
                3: "Renewable corn oil price",
                4: "Distillers grains price",
                5: "Approximate coproduct credit ($/bushel)",
                6: "Approximate coproduct credit ($/gal)",
                7: "Approximate coproduct credit ($m)",
                8: "Resolved source mode",
                9: "Coverage",
                10: "Rule",
            }
            for cc in range(frame_start_col, frame_end_col + 1):
                header_cell = ws.cell(row=frame_header_row, column=cc, value=frame_headers.get(cc, ""))
                header_cell.fill = copy(header_fill)
                header_cell.font = copy(bold_font)
                header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                header_cell.border = copy(thin_border)
            ws.row_dimensions[frame_header_row].height = 36.0

            frame_row = frame_header_row + 1
            frame_rows_by_key: Dict[str, int] = {}
            for rec in coproduct_frame_summary_records:
                row_fill = copy(zebra_fill_light if ((frame_row - frame_header_row) % 2) else zebra_fill_dark)
                row_values = {
                    2: str(rec.get("frame_label") or ""),
                    3: rec.get("renewable_corn_oil_price"),
                    4: rec.get("distillers_grains_price"),
                    5: rec.get("approximate_coproduct_credit"),
                    6: rec.get("approximate_coproduct_credit_per_gal"),
                    7: rec.get("approximate_coproduct_credit_usd_m"),
                    8: str(rec.get("resolved_source_mode") or "Unknown/blank"),
                    9: rec.get("coverage_ratio"),
                    10: str(rec.get("rule") or ""),
                }
                for cc in range(frame_start_col, frame_end_col + 1):
                    cell = ws.cell(row=frame_row, column=cc, value=row_values.get(cc))
                    cell.fill = copy(row_fill)
                    cell.font = copy(body_font)
                    cell.border = copy(thin_border)
                    cell.alignment = Alignment(horizontal="center" if cc in {2, 3, 4, 5, 6, 7, 9} else "left", vertical="center", wrap_text=cc in {8, 10})
                    if cc in {3, 4, 5, 6, 7} and pd.notna(pd.to_numeric(row_values.get(cc), errors="coerce")):
                        cell.number_format = "#,##0.000"
                    if cc == 9 and pd.notna(pd.to_numeric(row_values.get(cc), errors="coerce")):
                        cell.number_format = "0%"
                ws.row_dimensions[frame_row].height = 34.0
                frame_rows_by_key[str(rec.get("frame_key") or "")] = frame_row
                frame_row += 1
            coproduct_frame_summary_layout = {
                "title_row": frame_section_row,
                "note_row": frame_note_row,
                "header_row": frame_header_row,
                "frame_rows": frame_rows_by_key,
                "renewable_corn_oil_col": 3,
                "distillers_grains_col": 4,
                "approximate_coproduct_credit_col": 5,
                "approximate_coproduct_credit_per_gal_col": 6,
                "approximate_coproduct_credit_usd_m_col": 7,
                "resolved_source_mode_col": 8,
                "coverage_col": 9,
                "rule_col": 10,
                "next_row": frame_row - 1,
            }
        if coproduct_history_records:
            coproduct_history_section_row = int(
                coproduct_frame_summary_layout.get("next_row")
                or coproduct_signal_readiness_layout.get("next_row")
                or (coproduct_row - 1)
            ) + 2
            history_start_col = 2  # B
            history_end_col = 9  # I
            history_note_row = coproduct_history_section_row + 1
            history_header_row = coproduct_history_section_row + 2
            ws.merge_cells(
                start_row=coproduct_history_section_row,
                start_column=history_start_col,
                end_row=coproduct_history_section_row,
                end_column=history_end_col,
            )
            history_title = ws.cell(row=coproduct_history_section_row, column=history_start_col, value="Coproduct quarterly history")
            history_title.fill = copy(section_fill)
            history_title.font = copy(bold_font)
            history_title.alignment = Alignment(horizontal="center", vertical="center")
            history_title.border = copy(thin_border)
            for cc in range(history_start_col, history_end_col + 1):
                ws.cell(row=coproduct_history_section_row, column=cc).fill = copy(section_fill)
                ws.cell(row=coproduct_history_section_row, column=cc).font = copy(bold_font)
                ws.cell(row=coproduct_history_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=coproduct_history_section_row, column=cc).border = copy(thin_border)
            ws.row_dimensions[coproduct_history_section_row].height = 22.0

            ws.merge_cells(
                start_row=history_note_row,
                start_column=history_start_col,
                end_row=history_note_row,
                end_column=history_end_col,
            )
            history_note = ws.cell(
                row=history_note_row,
                column=history_start_col,
                value=(
                    "Historical price legs use quarter-aware active-capacity weighting across the active GPRE footprint. "
                    "DDGS uses all active plants, while corn oil uses the same active-footprint weighting as a producer-subset approximation because exact plant-level corn-oil coverage is not available. "
                    "$/gal = $/bushel divided by ethanol yield, coverage is covered active-capacity share, and the visible chart stays focused on approximate coproduct credit ($/gal) to avoid mixing price units."
                ),
            )
            history_note.fill = copy(intro_fill)
            history_note.font = copy(body_font)
            history_note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            history_note.border = copy(thin_border)
            for cc in range(history_start_col, history_end_col + 1):
                ws.cell(row=history_note_row, column=cc).fill = copy(intro_fill)
                ws.cell(row=history_note_row, column=cc).border = copy(thin_border)
            ws.row_dimensions[history_note_row].height = 52.0

            history_headers = {
                2: "Quarter",
                3: "Renewable corn oil price",
                4: "Distillers grains price",
                5: "Approximate coproduct credit ($/bushel)",
                6: "Approximate coproduct credit ($/gal)",
                7: "Approximate coproduct credit ($m)",
                8: "Resolved source mode",
                9: "Coverage",
            }
            for cc in range(history_start_col, history_end_col + 1):
                header_cell = ws.cell(row=history_header_row, column=cc, value=history_headers.get(cc, ""))
                header_cell.fill = copy(header_fill)
                header_cell.font = copy(bold_font)
                header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                header_cell.border = copy(thin_border)
            ws.row_dimensions[history_header_row].height = 30.0

            history_row = history_header_row + 1
            first_history_data_row = history_row
            for rec in coproduct_history_records:
                row_fill = copy(zebra_fill_light if ((history_row - history_header_row) % 2) else zebra_fill_dark)
                row_values = {
                    2: str(rec.get("quarter_label") or ""),
                    3: rec.get("renewable_corn_oil_price"),
                    4: rec.get("distillers_grains_price"),
                    5: rec.get("approximate_coproduct_credit"),
                    6: rec.get("approximate_coproduct_credit_per_gal"),
                    7: rec.get("approximate_coproduct_credit_usd_m"),
                    8: str(rec.get("resolved_source_mode") or "Unknown/blank"),
                    9: rec.get("coverage_ratio"),
                }
                for cc in range(history_start_col, history_end_col + 1):
                    cell = ws.cell(row=history_row, column=cc, value=row_values.get(cc))
                    cell.fill = copy(row_fill)
                    cell.font = copy(body_font)
                    cell.border = copy(thin_border)
                    cell.alignment = Alignment(horizontal="center" if cc in {2, 3, 4, 5, 6, 7, 9} else "left", vertical="center", wrap_text=cc == 8)
                    if cc in {3, 4, 5, 6, 7} and pd.notna(pd.to_numeric(row_values.get(cc), errors="coerce")):
                        cell.number_format = "#,##0.000"
                    if cc == 9 and pd.notna(pd.to_numeric(row_values.get(cc), errors="coerce")):
                        cell.number_format = "0%"
                ws.row_dimensions[history_row].height = 24.0
                history_row += 1

            coproduct_history_layout = {
                "title_row": coproduct_history_section_row,
                "note_row": history_note_row,
                "header_row": history_header_row,
                "first_data_row": first_history_data_row,
                "last_data_row": history_row - 1,
                "quarter_col": 2,
                "renewable_corn_oil_col": 3,
                "distillers_grains_col": 4,
                "approximate_coproduct_credit_col": 5,
                "approximate_coproduct_credit_per_gal_col": 6,
                "approximate_coproduct_credit_usd_m_col": 7,
                "resolved_source_mode_col": 8,
                "coverage_col": 9,
                "records": [dict(rec) for rec in list(coproduct_history_records or [])],
                "next_row": history_row - 1,
            }
        coproduct_volume_support_records = _coproduct_volume_support_audit_records()
        if coproduct_volume_support_records:
            volume_section_row = int(
                coproduct_history_layout.get("next_row")
                or coproduct_frame_summary_layout.get("next_row")
                or coproduct_signal_readiness_layout.get("next_row")
                or (coproduct_row - 1)
            ) + 2
            volume_note_row = volume_section_row + 1
            volume_header_row = volume_section_row + 2
            volume_start_col = 2  # B
            volume_end_col = 15  # O
            volume_spans = [
                (2, 3, "Series"),
                (4, 5, "Source/path"),
                (6, 6, "Historical usable"),
                (7, 7, "Current usable"),
                (8, 8, "Next usable"),
                (9, 10, "Best use"),
                (11, 15, "Note"),
            ]
            ws.merge_cells(start_row=volume_section_row, start_column=volume_start_col, end_row=volume_section_row, end_column=volume_end_col)
            volume_title = ws.cell(row=volume_section_row, column=volume_start_col, value="Coproduct volume support audit")
            volume_title.fill = copy(section_fill)
            volume_title.font = copy(bold_font)
            volume_title.alignment = Alignment(horizontal="center", vertical="center")
            volume_title.border = copy(thin_border)
            for cc in range(volume_start_col, volume_end_col + 1):
                ws.cell(row=volume_section_row, column=cc).fill = copy(section_fill)
                ws.cell(row=volume_section_row, column=cc).font = copy(bold_font)
                ws.cell(row=volume_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=volume_section_row, column=cc).border = copy(thin_border)
            ws.row_dimensions[volume_section_row].height = 22.0

            ws.merge_cells(start_row=volume_note_row, start_column=volume_start_col, end_row=volume_note_row, end_column=volume_end_col)
            volume_note = ws.cell(
                row=volume_note_row,
                column=volume_start_col,
                value=(
                    "Operating_Drivers coproduct volumes remain sourced historical actuals. "
                    "The latest actual volume/corn intensity is used only as a forward yield and $m anchor; "
                    "price coverage still comes from NWER/AMS market rows."
                ),
            )
            volume_note.fill = copy(intro_fill)
            volume_note.font = copy(body_font)
            volume_note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            volume_note.border = copy(thin_border)
            for cc in range(volume_start_col, volume_end_col + 1):
                ws.cell(row=volume_note_row, column=cc).fill = copy(intro_fill)
                ws.cell(row=volume_note_row, column=cc).border = copy(thin_border)
            ws.row_dimensions[volume_note_row].height = 32.0

            for start_col, end_col, header_txt in volume_spans:
                if end_col > start_col:
                    ws.merge_cells(start_row=volume_header_row, start_column=start_col, end_row=volume_header_row, end_column=end_col)
                for cc in range(start_col, end_col + 1):
                    ws.cell(row=volume_header_row, column=cc).fill = copy(header_fill)
                    ws.cell(row=volume_header_row, column=cc).font = copy(bold_font)
                    ws.cell(row=volume_header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws.cell(row=volume_header_row, column=cc).border = copy(thin_border)
                ws.cell(row=volume_header_row, column=start_col, value=header_txt)
            ws.row_dimensions[volume_header_row].height = 28.0

            volume_row = volume_header_row + 1
            volume_rows_by_key: Dict[str, int] = {}
            for rec in coproduct_volume_support_records:
                row_fill = copy(zebra_fill_light if ((volume_row - volume_header_row) % 2) else zebra_fill_dark)
                row_values = {
                    2: str(rec.get("label") or ""),
                    4: str(rec.get("source_path") or ""),
                    6: str(rec.get("historical_usable") or ""),
                    7: str(rec.get("current_usable") or ""),
                    8: str(rec.get("next_usable") or ""),
                    9: str(rec.get("best_use") or ""),
                    11: str(rec.get("note") or ""),
                }
                for start_col, end_col, _header_txt in volume_spans:
                    if end_col > start_col:
                        ws.merge_cells(start_row=volume_row, start_column=start_col, end_row=volume_row, end_column=end_col)
                    for cc in range(start_col, end_col + 1):
                        ws.cell(row=volume_row, column=cc).fill = copy(row_fill)
                        ws.cell(row=volume_row, column=cc).font = copy(body_font)
                        ws.cell(row=volume_row, column=cc).border = copy(thin_border)
                        ws.cell(row=volume_row, column=cc).alignment = Alignment(
                            horizontal="center" if start_col in {6, 7, 8, 9} else "left",
                            vertical="top",
                            wrap_text=True,
                        )
                    ws.cell(row=volume_row, column=start_col, value=row_values.get(start_col, ""))
                ws.row_dimensions[volume_row].height = 34.0
                volume_rows_by_key[str(rec.get("driver_key") or "")] = volume_row
                volume_row += 1
            coproduct_volume_support_layout = {
                "title_row": volume_section_row,
                "note_row": volume_note_row,
                "header_row": volume_header_row,
                "rows_by_key": volume_rows_by_key,
                "next_row": volume_row - 1,
            }
        if (
            not coproduct_experimental_candidate_comparison_df.empty
            or best_coproduct_experimental_model_key
        ):
            coproduct_experimental_specs_by_key = {
                str(rec.get("model_key") or ""): dict(rec)
                for rec in coproduct_experimental_method_specs
                if str(rec.get("model_key") or "").strip()
            }

            def _coproduct_experimental_frame_value(model_key_in: Any, frame_key_in: str) -> Optional[float]:
                model_frames = coproduct_experimental_frame_values.get(str(model_key_in or "")) if isinstance(coproduct_experimental_frame_values, dict) else {}
                frame = model_frames.get(str(frame_key_in or "")) if isinstance(model_frames, dict) else {}
                value_num = pd.to_numeric((frame or {}).get("value_usd_per_gal"), errors="coerce")
                if pd.isna(value_num):
                    return None
                return float(value_num)

            experimental_section_row = int(
                coproduct_volume_support_layout.get("next_row")
                or coproduct_history_layout.get("next_row")
                or coproduct_frame_summary_layout.get("next_row")
                or coproduct_signal_readiness_layout.get("next_row")
                or (coproduct_row - 1)
            ) + 2
            experimental_note_row = experimental_section_row + 1
            experimental_start_col = 2  # B
            experimental_end_col = 15  # O
            ws.merge_cells(
                start_row=experimental_section_row,
                start_column=experimental_start_col,
                end_row=experimental_section_row,
                end_column=experimental_end_col,
            )
            experimental_title = ws.cell(
                row=experimental_section_row,
                column=experimental_start_col,
                value="Coproduct-aware experimental lenses",
            )
            experimental_title.fill = copy(section_fill)
            experimental_title.font = copy(bold_font)
            experimental_title.alignment = Alignment(horizontal="center", vertical="center")
            experimental_title.border = copy(thin_border)
            for cc in range(experimental_start_col, experimental_end_col + 1):
                ws.cell(row=experimental_section_row, column=cc).fill = copy(section_fill)
                ws.cell(row=experimental_section_row, column=cc).font = copy(bold_font)
                ws.cell(row=experimental_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=experimental_section_row, column=cc).border = copy(thin_border)
            ws.row_dimensions[experimental_section_row].height = 22.0

            ws.merge_cells(
                start_row=experimental_note_row,
                start_column=experimental_start_col,
                end_row=experimental_note_row,
                end_column=experimental_end_col,
            )
            experimental_note = ws.cell(
                row=experimental_note_row,
                column=experimental_start_col,
                value=(
                    "Comparison only. Coproduct-aware overlays are sandbox lenses and do not compete for production promotion."
                ),
            )
            experimental_note.fill = copy(intro_fill)
            experimental_note.font = copy(body_font)
            experimental_note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            experimental_note.border = copy(thin_border)
            for cc in range(experimental_start_col, experimental_end_col + 1):
                ws.cell(row=experimental_note_row, column=cc).fill = copy(intro_fill)
                ws.cell(row=experimental_note_row, column=cc).border = copy(thin_border)
            ws.row_dimensions[experimental_note_row].height = 30.0

            summary_header_row = experimental_note_row + 1
            ws.cell(row=summary_header_row, column=2, value="Summary").fill = copy(header_fill)
            ws.cell(row=summary_header_row, column=2).font = copy(bold_font)
            ws.cell(row=summary_header_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=summary_header_row, column=2).border = copy(thin_border)
            for cc in range(3, 16):
                ws.cell(row=summary_header_row, column=cc).fill = copy(header_fill)
                ws.cell(row=summary_header_row, column=cc).font = copy(bold_font)
                ws.cell(row=summary_header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=summary_header_row, column=cc).border = copy(thin_border)
            ws.merge_cells(start_row=summary_header_row, start_column=3, end_row=summary_header_row, end_column=15)
            ws.cell(row=summary_header_row, column=3, value="Value")
            ws.row_dimensions[summary_header_row].height = 22.0

            best_coproduct_spec = dict(
                coproduct_experimental_specs_by_key.get(best_coproduct_experimental_model_key) or {}
            )
            previous_coproduct_reference_label = (
                _sandbox_model_label(coproduct_experimental_legacy_reference_model_key)
                or str(coproduct_experimental_legacy_reference_row.get("method_label") or "")
                or "Simple + 50% credit"
            )
            summary_rows = [
                ("Best coproduct-aware experimental lens", _sandbox_model_label(best_coproduct_experimental_model_key)),
                ("Definition", str(best_coproduct_spec.get("rule") or "")),
                ("Prior quarter ($/gal)", _coproduct_experimental_frame_value(best_coproduct_experimental_model_key, "prior_quarter")),
                ("Quarter-open outlook ($/gal)", _coproduct_experimental_frame_value(best_coproduct_experimental_model_key, "quarter_open")),
                ("Current QTD ($/gal)", _coproduct_experimental_frame_value(best_coproduct_experimental_model_key, "current_qtd")),
                ("Next quarter outlook ($/gal)", _coproduct_experimental_frame_value(best_coproduct_experimental_model_key, "next_quarter_thesis")),
                ("Best historical coproduct-aware", _sandbox_model_label(best_coproduct_experimental_historical_model_key)),
                ("Best forward coproduct-aware", _sandbox_model_label(best_coproduct_experimental_forward_model_key)),
                ("Previous best coproduct-aware (reference)", previous_coproduct_reference_label),
                ("Current production winner (reference)", _sandbox_model_label(production_winner_model_key)),
                ("Promotion status", "Experimental only"),
            ]
            summary_row = summary_header_row + 1
            summary_rows_by_label: Dict[str, int] = {}
            for label_txt, value_obj in summary_rows:
                row_fill = copy(zebra_fill_light if ((summary_row - summary_header_row) % 2) else zebra_fill_dark)
                label_cell = ws.cell(row=summary_row, column=2, value=label_txt)
                label_cell.fill = copy(row_fill)
                label_cell.font = copy(body_font)
                label_cell.border = copy(thin_border)
                label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                for cc in range(3, 16):
                    ws.cell(row=summary_row, column=cc).fill = copy(row_fill)
                    ws.cell(row=summary_row, column=cc).border = copy(thin_border)
                    ws.cell(row=summary_row, column=cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws.merge_cells(start_row=summary_row, start_column=3, end_row=summary_row, end_column=15)
                value_cell = ws.cell(row=summary_row, column=3, value=value_obj)
                value_cell.fill = copy(row_fill)
                value_cell.font = copy(body_font)
                value_cell.border = copy(thin_border)
                value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if label_txt.endswith("($/gal)") and pd.notna(pd.to_numeric(value_obj, errors="coerce")):
                    value_cell.number_format = "#,##0.000"
                ws.row_dimensions[summary_row].height = 28.0 if label_txt != "Definition" else 34.0
                summary_rows_by_label[label_txt] = summary_row
                summary_row += 1

            comparison_header_row = summary_row + 1
            comparison_headers = {
                2: "Method",
                3: "Rule",
                4: "Clean MAE",
                5: "Hybrid",
                6: "Hard-Q MAE",
                7: "Sign accuracy",
                8: "Forward",
                9: "Complexity",
                10: "Low-coverage MAE",
                11: "Δ vs current winner",
                12: "Status",
            }
            for cc in range(2, 13):
                header_cell = ws.cell(row=comparison_header_row, column=cc, value=comparison_headers.get(cc, ""))
                header_cell.fill = copy(header_fill)
                header_cell.font = copy(bold_font)
                header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                header_cell.border = copy(thin_border)
            ws.row_dimensions[comparison_header_row].height = 34.0

            comparison_rows_by_key: Dict[str, int] = {}
            comparison_row = comparison_header_row + 1
            sorted_coproduct_candidates = coproduct_experimental_candidate_comparison_df.sort_values(
                ["hybrid_score", "clean_window_mae"],
                na_position="last",
            )
            for rec in sorted_coproduct_candidates.to_dict("records"):
                row_fill = copy(zebra_fill_light if ((comparison_row - comparison_header_row) % 2) else zebra_fill_dark)
                row_values = {
                    2: _sandbox_model_label(rec.get("model_key")),
                    3: str(rec.get("rule") or ""),
                    4: rec.get("clean_window_mae"),
                    5: rec.get("hybrid_score"),
                    6: rec.get("hard_quarter_mae"),
                    7: rec.get("sign_accuracy"),
                    8: str(rec.get("forward_usability_rating") or ""),
                    9: str(rec.get("complexity_rating") or ""),
                    10: rec.get("low_coverage_mae"),
                    11: rec.get("delta_vs_current_winner"),
                    12: str(rec.get("status") or "comparison only"),
                }
                for cc in range(2, 13):
                    cell = ws.cell(row=comparison_row, column=cc, value=row_values.get(cc))
                    cell.fill = copy(row_fill)
                    cell.font = copy(body_font)
                    cell.border = copy(thin_border)
                    cell.alignment = Alignment(horizontal="center" if cc != 3 else "left", vertical="center", wrap_text=cc in {2, 3, 12})
                    if cc in {4, 5, 6, 10, 11} and pd.notna(pd.to_numeric(row_values.get(cc), errors="coerce")):
                        cell.number_format = "#,##0.000"
                    if cc == 7 and pd.notna(pd.to_numeric(row_values.get(cc), errors="coerce")):
                        cell.number_format = "0.0%"
                ws.row_dimensions[comparison_row].height = 32.0
                comparison_rows_by_key[str(rec.get("model_key") or "")] = comparison_row
                comparison_row += 1

            coproduct_experimental_layout = {
                "title_row": experimental_section_row,
                "note_row": experimental_note_row,
                "summary_header_row": summary_header_row,
                "summary_rows_by_label": summary_rows_by_label,
                "comparison_header_row": comparison_header_row,
                "comparison_rows_by_key": comparison_rows_by_key,
                "next_row": comparison_row - 1,
                "summary_markdown": coproduct_experimental_summary_md,
            }
    memo_section_row = int(
        coproduct_experimental_layout.get("next_row")
        or coproduct_volume_support_layout.get("next_row")
        or coproduct_history_layout.get("next_row")
        or coproduct_frame_summary_layout.get("next_row")
        or coproduct_signal_readiness_layout.get("next_row")
        or approx_market_crush_build_up_layout.get("next_row")
        or process_build_section_row
    ) + 2
    ws.merge_cells(start_row=memo_section_row, start_column=2, end_row=memo_section_row, end_column=15)
    memo_title = ws.cell(row=memo_section_row, column=2, value="Hedge-adjusted memo tests")
    memo_title.fill = copy(section_fill)
    memo_title.font = copy(bold_font)
    memo_title.alignment = Alignment(horizontal="center", vertical="center")
    memo_title.border = copy(thin_border)
    for cc in range(2, 16):
        ws.cell(row=memo_section_row, column=cc).fill = copy(section_fill)
        ws.cell(row=memo_section_row, column=cc).font = copy(bold_font)
        ws.cell(row=memo_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=memo_section_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[memo_section_row].height = 22.0

    memo_header_row = memo_section_row + 1
    memo_headers = [
        "Quarter",
        "Target type",
        "Target $/gal",
        "Simple market",
        "GPRE proxy",
        "Disclosed hedge %",
        "Pattern hedge %",
        "Memo disclosed bridge prior-qtr",
        "Memo disclosed process prior-qtr",
        "Memo pattern bridge prior-qtr",
        "Memo pattern process prior-qtr",
        "Comment",
    ]
    for offset, header in enumerate(memo_headers, start=2):
        cell = ws.cell(row=memo_header_row, column=offset, value=header)
        cell.fill = copy(header_fill)
        cell.font = copy(bold_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(thin_border)
    ws.row_dimensions[memo_header_row].height = 34.0

    memo_cols = [
        "quarter_label",
        "target_basis",
        "evaluation_target_margin_usd_per_gal",
        "simple_market_proxy_usd_per_gal",
        "gpre_proxy_official_usd_per_gal",
        "hedge_share_disclosed",
        "hedge_share_pattern",
        "hedge_memo_disclosed_bridge_prior_current_usd_per_gal",
        "hedge_memo_disclosed_process_prior_current_usd_per_gal",
        "hedge_memo_pattern_bridge_prior_current_usd_per_gal",
        "hedge_memo_pattern_process_prior_current_usd_per_gal",
    ]
    memo_row = memo_header_row + 1
    if isinstance(quarterly_df, pd.DataFrame) and not quarterly_df.empty:
        for rec in quarterly_df.to_dict("records"):
            quarter_txt = str(rec.get("quarter_label") or "")
            if not quarter_txt.startswith("2025-"):
                continue
            comment_bits = []
            disclosed_share = pd.to_numeric(rec.get("hedge_share_disclosed"), errors="coerce")
            pattern_share = pd.to_numeric(rec.get("hedge_share_pattern"), errors="coerce")
            if pd.notna(disclosed_share) and float(disclosed_share) > 0:
                comment_bits.append("Uses explicit disclosed hedge share")
            if pd.notna(pattern_share) and float(pattern_share) > 0:
                comment_bits.append("Pattern memo uses fixed seasonal hedge share")
            else:
                comment_bits.append("No hedge signal; spot bridge only")
            vals = [rec.get(col_name) for col_name in memo_cols] + [" | ".join(comment_bits)]
            for offset, val in enumerate(vals, start=2):
                if offset == 3:
                    val = str(val or "").strip().title()
                cell = ws.cell(row=memo_row, column=offset, value=val)
                cell.fill = copy(zebra_fill_light if ((memo_row - memo_header_row) % 2) else zebra_fill_dark)
                cell.font = copy(body_font)
                cell.alignment = Alignment(horizontal="left" if offset in {2, 3, 13} else "center", vertical="top", wrap_text=offset in {13})
                cell.border = copy(thin_border)
                if offset in {4, 5, 6, 7, 8, 9, 10, 11, 12} and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0.000"
            ws.row_dimensions[memo_row].height = 30.0
            memo_row += 1

    hedge_section_row = memo_row + 1
    ws.merge_cells(start_row=hedge_section_row, start_column=2, end_row=hedge_section_row, end_column=15)
    hedge_title = ws.cell(row=hedge_section_row, column=2, value="Implied hedge / realization style study")
    hedge_title.fill = copy(section_fill)
    hedge_title.font = copy(bold_font)
    hedge_title.alignment = Alignment(horizontal="center", vertical="center")
    hedge_title.border = copy(thin_border)
    for cc in range(2, 16):
        ws.cell(row=hedge_section_row, column=cc).fill = copy(section_fill)
        ws.cell(row=hedge_section_row, column=cc).font = copy(bold_font)
        ws.cell(row=hedge_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=hedge_section_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[hedge_section_row].height = 22.0

    hedge_summary_row = hedge_section_row + 1
    ws.merge_cells(start_row=hedge_summary_row, start_column=2, end_row=hedge_summary_row, end_column=15)
    hedge_summary_text = (
        f"Target: {hedge_target_label}. "
        f"Usable quarters: {int(len(hedge_quarter_fit_df))}. "
        f"Backtest window: {hedge_backtest_window_display or 'n/a'}. "
        f"Best overall style: {hedge_best_style_display or 'n/a'}. "
        f"Best overall family: {hedge_best_family_display or 'n/a'}."
    )
    hedge_summary_cell = ws.cell(row=hedge_summary_row, column=2, value=hedge_summary_text)
    hedge_summary_cell.fill = copy(intro_fill)
    hedge_summary_cell.font = copy(body_font)
    hedge_summary_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    hedge_summary_cell.border = copy(thin_border)
    hedge_summary_comment_bits = [bit for bit in [hedge_target_definition, hedge_style_vs_family_explanation, hedge_diagnostic_only_note] if str(bit or "").strip()]
    if hedge_summary_comment_bits:
        hedge_summary_cell.comment = Comment("\n".join(hedge_summary_comment_bits), "Codex")
    for cc in range(2, 16):
        ws.cell(row=hedge_summary_row, column=cc).fill = copy(intro_fill)
        ws.cell(row=hedge_summary_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[hedge_summary_row].height = 32.0

    hedge_note_row = hedge_summary_row + 1
    ws.merge_cells(start_row=hedge_note_row, start_column=2, end_row=hedge_note_row, end_column=15)
    hedge_note_text = " ".join(
        bit
        for bit in [
            hedge_style_vs_family_explanation,
            hedge_diagnostic_only_note,
        ]
        if str(bit or "").strip()
    )
    hedge_note_cell = ws.cell(row=hedge_note_row, column=2, value=hedge_note_text or "Diagnostic only.")
    hedge_note_cell.fill = copy(intro_fill)
    hedge_note_cell.font = copy(body_font)
    hedge_note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    hedge_note_cell.border = copy(thin_border)
    for cc in range(2, 16):
        ws.cell(row=hedge_note_row, column=cc).fill = copy(intro_fill)
        ws.cell(row=hedge_note_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[hedge_note_row].height = 34.0

    hedge_leader_title_row = hedge_note_row + 2
    ws.merge_cells(start_row=hedge_leader_title_row, start_column=2, end_row=hedge_leader_title_row, end_column=8)
    hedge_leader_title = ws.cell(row=hedge_leader_title_row, column=2, value="Hedge-style family leaderboard")
    hedge_leader_title.fill = copy(section_fill)
    hedge_leader_title.font = copy(bold_font)
    hedge_leader_title.alignment = Alignment(horizontal="center", vertical="center")
    hedge_leader_title.border = copy(thin_border)
    for cc in range(2, 9):
        ws.cell(row=hedge_leader_title_row, column=cc).fill = copy(section_fill)
        ws.cell(row=hedge_leader_title_row, column=cc).font = copy(bold_font)
        ws.cell(row=hedge_leader_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=hedge_leader_title_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[hedge_leader_title_row].height = 22.0

    hedge_leader_header_row = hedge_leader_title_row + 1
    hedge_leader_headers = ["Style", "Family", "MAE", "Mean error", "Sign hit-rate", "Best-fit quarters"]
    for offset, header in enumerate(hedge_leader_headers, start=2):
        cell = ws.cell(row=hedge_leader_header_row, column=offset, value=header)
        cell.fill = copy(header_fill)
        cell.font = copy(bold_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(thin_border)
    ws.row_dimensions[hedge_leader_header_row].height = 30.0

    hedge_leader_row = hedge_leader_header_row + 1
    if isinstance(hedge_candidate_leaderboard_df, pd.DataFrame) and not hedge_candidate_leaderboard_df.empty:
        for rec in hedge_candidate_leaderboard_df.to_dict("records"):
            vals = [
                _sandbox_model_label(rec.get("style_key") or rec.get("style_label") or ""),
                str(rec.get("family_label") or rec.get("family") or "").strip(),
                rec.get("mae"),
                rec.get("mean_error"),
                (
                    None
                    if pd.isna(pd.to_numeric(rec.get("sign_hit_rate"), errors="coerce"))
                    else float(pd.to_numeric(rec.get("sign_hit_rate"), errors="coerce")) * 100.0
                ),
                rec.get("best_fit_quarter_count"),
            ]
            for offset, val in enumerate(vals, start=2):
                cell = ws.cell(row=hedge_leader_row, column=offset, value=val)
                cell.fill = copy(zebra_fill_light if ((hedge_leader_row - hedge_leader_header_row) % 2) else zebra_fill_dark)
                cell.font = copy(body_font)
                cell.alignment = Alignment(horizontal="left" if offset in {2, 3} else "center", vertical="top", wrap_text=offset in {2, 3})
                cell.border = copy(thin_border)
                if offset in {4, 5} and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0.000"
                if offset == 6 and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0.0"
            ws.row_dimensions[hedge_leader_row].height = 28.0
            hedge_leader_row += 1
    else:
        ws.merge_cells(start_row=hedge_leader_row, start_column=2, end_row=hedge_leader_row, end_column=8)
        hedge_leader_empty = ws.cell(row=hedge_leader_row, column=2, value="No usable reported-target quarters were available for the hedge-style study.")
        hedge_leader_empty.fill = copy(intro_fill)
        hedge_leader_empty.font = copy(body_font)
        hedge_leader_empty.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        hedge_leader_empty.border = copy(thin_border)
        for cc in range(2, 9):
            ws.cell(row=hedge_leader_row, column=cc).fill = copy(intro_fill)
            ws.cell(row=hedge_leader_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[hedge_leader_row].height = 28.0
        hedge_leader_row += 1

    hedge_quarter_title_row = hedge_leader_row + 1
    ws.merge_cells(start_row=hedge_quarter_title_row, start_column=2, end_row=hedge_quarter_title_row, end_column=9)
    hedge_quarter_title = ws.cell(row=hedge_quarter_title_row, column=2, value="Quarter-by-quarter best-fit hedge style")
    hedge_quarter_title.fill = copy(section_fill)
    hedge_quarter_title.font = copy(bold_font)
    hedge_quarter_title.alignment = Alignment(horizontal="center", vertical="center")
    hedge_quarter_title.border = copy(thin_border)
    for cc in range(2, 10):
        ws.cell(row=hedge_quarter_title_row, column=cc).fill = copy(section_fill)
        ws.cell(row=hedge_quarter_title_row, column=cc).font = copy(bold_font)
        ws.cell(row=hedge_quarter_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=hedge_quarter_title_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[hedge_quarter_title_row].height = 22.0

    hedge_quarter_header_row = hedge_quarter_title_row + 1
    hedge_quarter_headers = [
        "Quarter",
        "Reported consolidated crush margin ($/gal)",
        "Best-fit style",
        "Best-fit value",
        "Error",
        "Weak fit?",
        "Hard quarter?",
        "Note/category",
    ]
    for offset, header in enumerate(hedge_quarter_headers, start=2):
        cell = ws.cell(row=hedge_quarter_header_row, column=offset, value=header)
        cell.fill = copy(header_fill)
        cell.font = copy(bold_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(thin_border)
    ws.row_dimensions[hedge_quarter_header_row].height = 34.0

    hedge_quarter_row = hedge_quarter_header_row + 1
    if isinstance(hedge_quarter_fit_df, pd.DataFrame) and not hedge_quarter_fit_df.empty:
        for rec in hedge_quarter_fit_df.to_dict("records"):
            hard_reason = str(rec.get("hard_quarter_reason") or "").strip()
            hard_txt = "No"
            if bool(rec.get("hard_quarter_flag")):
                hard_txt = f"Yes: {hard_reason}" if hard_reason else "Yes"
            vals = [
                rec.get("quarter_label"),
                rec.get("target_value_usd_per_gal"),
                _sandbox_model_label(rec.get("best_fit_style_key") or rec.get("best_fit_style_label") or ""),
                rec.get("best_fit_value_usd_per_gal"),
                rec.get("best_fit_error_usd_per_gal"),
                rec.get("weak_fit_display") or ("Yes" if bool(rec.get("weak_fit_flag")) else "No"),
                hard_txt,
                rec.get("fit_note"),
            ]
            for offset, val in enumerate(vals, start=2):
                cell = ws.cell(row=hedge_quarter_row, column=offset, value=val)
                cell.fill = copy(zebra_fill_light if ((hedge_quarter_row - hedge_quarter_header_row) % 2) else zebra_fill_dark)
                cell.font = copy(body_font)
                cell.alignment = Alignment(horizontal="left" if offset in {2, 4, 8, 9} else "center", vertical="top", wrap_text=offset in {4, 8, 9})
                cell.border = copy(thin_border)
                if offset in {3, 5, 6} and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0.000"
            ws.row_dimensions[hedge_quarter_row].height = 32.0
            hedge_quarter_row += 1
    else:
        ws.merge_cells(start_row=hedge_quarter_row, start_column=2, end_row=hedge_quarter_row, end_column=9)
        hedge_quarter_empty = ws.cell(row=hedge_quarter_row, column=2, value="No best-fit quarter table was available for the hedge-style study.")
        hedge_quarter_empty.fill = copy(intro_fill)
        hedge_quarter_empty.font = copy(body_font)
        hedge_quarter_empty.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        hedge_quarter_empty.border = copy(thin_border)
        for cc in range(2, 10):
            ws.cell(row=hedge_quarter_row, column=cc).fill = copy(intro_fill)
            ws.cell(row=hedge_quarter_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[hedge_quarter_row].height = 26.0
        hedge_quarter_row += 1

    hedge_interp_title_row = hedge_quarter_row + 1
    ws.merge_cells(start_row=hedge_interp_title_row, start_column=2, end_row=hedge_interp_title_row, end_column=15)
    hedge_interp_title = ws.cell(row=hedge_interp_title_row, column=2, value="Interpretation")
    hedge_interp_title.fill = copy(section_fill)
    hedge_interp_title.font = copy(bold_font)
    hedge_interp_title.alignment = Alignment(horizontal="center", vertical="center")
    hedge_interp_title.border = copy(thin_border)
    for cc in range(2, 16):
        ws.cell(row=hedge_interp_title_row, column=cc).fill = copy(section_fill)
        ws.cell(row=hedge_interp_title_row, column=cc).font = copy(bold_font)
        ws.cell(row=hedge_interp_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=hedge_interp_title_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[hedge_interp_title_row].height = 22.0

    hedge_interp_row = hedge_interp_title_row + 1
    ws.merge_cells(start_row=hedge_interp_row, start_column=2, end_row=hedge_interp_row, end_column=15)
    hedge_interp_text = "\n".join(hedge_interpretation_lines) if hedge_interpretation_lines else "No additional hedge-style interpretation was available."
    hedge_interp_cell = ws.cell(row=hedge_interp_row, column=2, value=hedge_interp_text)
    hedge_interp_cell.fill = copy(intro_fill)
    hedge_interp_cell.font = copy(body_font)
    hedge_interp_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    hedge_interp_cell.border = copy(thin_border)
    for cc in range(2, 16):
        ws.cell(row=hedge_interp_row, column=cc).fill = copy(intro_fill)
        ws.cell(row=hedge_interp_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[hedge_interp_row].height = 42.0

    futures_section_row = hedge_interp_row + 2
    ws.merge_cells(start_row=futures_section_row, start_column=2, end_row=futures_section_row, end_column=20)
    futures_title = ws.cell(row=futures_section_row, column=2, value="Futures timing / hedge-style sandbox")
    futures_title.fill = copy(section_fill)
    futures_title.font = copy(bold_font)
    futures_title.alignment = Alignment(horizontal="center", vertical="center")
    futures_title.border = copy(thin_border)
    for cc in range(2, 21):
        ws.cell(row=futures_section_row, column=cc).fill = copy(section_fill)
        ws.cell(row=futures_section_row, column=cc).font = copy(bold_font)
        ws.cell(row=futures_section_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=futures_section_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[futures_section_row].height = 24.0

    futures_note_row = futures_section_row + 1
    ws.merge_cells(start_row=futures_note_row, start_column=2, end_row=futures_note_row, end_column=20)
    futures_note_bits = [
        futures_timing_diagnostic_note or "Sandbox/comparison-only; does not change Economics_Overlay, Approximate market crush, or GPRE crush proxy.",
        f"Target: {futures_timing_target_label}.",
    ]
    if futures_timing_coverage_note:
        futures_note_bits.append(futures_timing_coverage_note)
    futures_note_cell = ws.cell(row=futures_note_row, column=2, value=" ".join(futures_note_bits))
    futures_note_cell.fill = copy(intro_fill)
    futures_note_cell.font = copy(body_font)
    futures_note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    futures_note_cell.border = copy(thin_border)
    for cc in range(2, 21):
        ws.cell(row=futures_note_row, column=cc).fill = copy(intro_fill)
        ws.cell(row=futures_note_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[futures_note_row].height = 42.0

    futures_summary_title_row = futures_note_row + 1
    ws.merge_cells(start_row=futures_summary_title_row, start_column=2, end_row=futures_summary_title_row, end_column=15)
    futures_summary_title = ws.cell(row=futures_summary_title_row, column=2, value="Futures timing candidate summary")
    futures_summary_title.fill = copy(section_fill)
    futures_summary_title.font = copy(bold_font)
    futures_summary_title.alignment = Alignment(horizontal="center", vertical="center")
    futures_summary_title.border = copy(thin_border)
    for cc in range(2, 16):
        ws.cell(row=futures_summary_title_row, column=cc).fill = copy(section_fill)
        ws.cell(row=futures_summary_title_row, column=cc).font = copy(bold_font)
        ws.cell(row=futures_summary_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=futures_summary_title_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[futures_summary_title_row].height = 22.0

    futures_summary_header_row = futures_summary_title_row + 1
    futures_summary_headers = [
        "Candidate",
        "Family",
        "Timing window",
        "Weighting style",
        "Timing rule",
        "Locked commodities",
        "Usable quarters",
        "MAE",
        "Median abs err",
        "Max err",
        "Closest qtrs",
        "Coverage %",
        "Status",
        "Notes",
    ]
    for offset, header in enumerate(futures_summary_headers, start=2):
        cell = ws.cell(row=futures_summary_header_row, column=offset, value=header)
        cell.fill = copy(header_fill)
        cell.font = copy(bold_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(thin_border)
    ws.row_dimensions[futures_summary_header_row].height = 34.0

    futures_summary_row = futures_summary_header_row + 1
    if isinstance(futures_timing_leaderboard_df, pd.DataFrame) and not futures_timing_leaderboard_df.empty:
        for rec in futures_timing_leaderboard_df.to_dict("records"):
            vals = [
                rec.get("candidate_label") or rec.get("candidate_key"),
                rec.get("family_label") or rec.get("family"),
                rec.get("timing_window"),
                rec.get("weighting_style"),
                rec.get("timing_rule"),
                rec.get("locked_commodities_label"),
                rec.get("usable_quarter_count"),
                rec.get("mae"),
                rec.get("median_abs_error"),
                rec.get("max_abs_error"),
                rec.get("closest_quarter_count"),
                rec.get("avg_anchor_coverage_ratio"),
                rec.get("status"),
                rec.get("notes"),
            ]
            for offset, val in enumerate(vals, start=2):
                cell = ws.cell(row=futures_summary_row, column=offset, value=val)
                cell.fill = copy(zebra_fill_light if ((futures_summary_row - futures_summary_header_row) % 2) else zebra_fill_dark)
                cell.font = copy(body_font)
                cell.alignment = Alignment(horizontal="left" if offset in {2, 3, 4, 5, 6, 7, 14, 15} else "center", vertical="top", wrap_text=offset in {2, 3, 4, 5, 6, 7, 14, 15})
                cell.border = copy(thin_border)
                if offset in {9, 10, 11} and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0.000"
                if offset == 13 and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0%"
            ws.row_dimensions[futures_summary_row].height = 32.0
            futures_summary_row += 1
    else:
        ws.merge_cells(start_row=futures_summary_row, start_column=2, end_row=futures_summary_row, end_column=15)
        cell = ws.cell(row=futures_summary_row, column=2, value="No futures timing sandbox candidates were available.")
        cell.fill = copy(intro_fill)
        cell.font = copy(body_font)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = copy(thin_border)
        for cc in range(2, 16):
            ws.cell(row=futures_summary_row, column=cc).fill = copy(intro_fill)
            ws.cell(row=futures_summary_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[futures_summary_row].height = 26.0
        futures_summary_row += 1

    futures_detail_title_row = futures_summary_row + 1
    ws.merge_cells(start_row=futures_detail_title_row, start_column=2, end_row=futures_detail_title_row, end_column=20)
    futures_detail_title = ws.cell(row=futures_detail_title_row, column=2, value="Futures timing quarter detail")
    futures_detail_title.fill = copy(section_fill)
    futures_detail_title.font = copy(bold_font)
    futures_detail_title.alignment = Alignment(horizontal="center", vertical="center")
    futures_detail_title.border = copy(thin_border)
    for cc in range(2, 21):
        ws.cell(row=futures_detail_title_row, column=cc).fill = copy(section_fill)
        ws.cell(row=futures_detail_title_row, column=cc).font = copy(bold_font)
        ws.cell(row=futures_detail_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=futures_detail_title_row, column=cc).border = copy(thin_border)
    ws.row_dimensions[futures_detail_title_row].height = 22.0

    futures_detail_header_row = futures_detail_title_row + 1
    futures_detail_headers = [
        "Quarter",
        "Candidate",
        "Target",
        "Official",
        "GPRE proxy",
        "Best forward",
        "Prediction",
        "Error",
        "Timing window",
        "Weighting style",
        "Anchors",
        "Coverage %",
        "Anchor count",
        "Status",
        "Symbols",
        "Obs dates",
        "Prices",
        "Source files",
        "Basis / missing",
    ]
    for offset, header in enumerate(futures_detail_headers, start=2):
        cell = ws.cell(row=futures_detail_header_row, column=offset, value=header)
        cell.fill = copy(header_fill)
        cell.font = copy(bold_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(thin_border)
    ws.row_dimensions[futures_detail_header_row].height = 34.0

    futures_detail_row = futures_detail_header_row + 1
    if isinstance(futures_timing_detail_df, pd.DataFrame) and not futures_timing_detail_df.empty:
        for rec in futures_timing_detail_df.to_dict("records"):
            expected_anchor_count = pd.to_numeric(rec.get("expected_anchor_count"), errors="coerce")
            usable_anchor_count = pd.to_numeric(rec.get("usable_anchor_count"), errors="coerce")
            anchor_count_text = (
                ""
                if pd.isna(expected_anchor_count) or pd.isna(usable_anchor_count)
                else f"{int(usable_anchor_count)}/{int(expected_anchor_count)}"
            )
            vals = [
                rec.get("quarter_label"),
                rec.get("candidate_label") or rec.get("candidate_key"),
                rec.get("target_value_usd_per_gal"),
                rec.get("official_simple_proxy_usd_per_gal"),
                rec.get("gpre_proxy_official_usd_per_gal"),
                rec.get("best_forward_lens_proxy_usd_per_gal"),
                rec.get("pred_value_usd_per_gal"),
                rec.get("error_usd_per_gal"),
                rec.get("timing_window"),
                rec.get("weighting_style"),
                rec.get("anchor_dates"),
                rec.get("anchor_coverage_ratio"),
                anchor_count_text,
                rec.get("availability_status"),
                rec.get("selected_symbols"),
                rec.get("selected_observation_dates"),
                rec.get("selected_prices"),
                rec.get("source_files"),
                " | ".join(
                    item
                    for item in [
                        str(rec.get("basis_source") or "").strip(),
                        str(rec.get("missing_data_flags") or "").strip(),
                    ]
                    if item
                ),
            ]
            for offset, val in enumerate(vals, start=2):
                cell = ws.cell(row=futures_detail_row, column=offset, value=val)
                cell.fill = copy(zebra_fill_light if ((futures_detail_row - futures_detail_header_row) % 2) else zebra_fill_dark)
                cell.font = copy(body_font)
                cell.alignment = Alignment(horizontal="left" if offset in {3, 10, 11, 12, 15, 16, 17, 18, 19, 20} else "center", vertical="top", wrap_text=offset in {3, 10, 11, 12, 15, 16, 17, 18, 19, 20})
                cell.border = copy(thin_border)
                if offset in {4, 5, 6, 7, 8, 9} and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0.000"
                if offset == 13 and pd.notna(pd.to_numeric(val, errors="coerce")):
                    cell.number_format = "0%"
            ws.row_dimensions[futures_detail_row].height = 34.0
            futures_detail_row += 1
    else:
        ws.merge_cells(start_row=futures_detail_row, start_column=2, end_row=futures_detail_row, end_column=20)
        cell = ws.cell(row=futures_detail_row, column=2, value="No futures timing quarter detail rows were available.")
        cell.fill = copy(intro_fill)
        cell.font = copy(body_font)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = copy(thin_border)
        for cc in range(2, 21):
            ws.cell(row=futures_detail_row, column=cc).fill = copy(intro_fill)
            ws.cell(row=futures_detail_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[futures_detail_row].height = 26.0
        futures_detail_row += 1
    futures_detail_end_row = futures_detail_row - 1

    _style_note_box_row(intro_row, 2, 15, fill=note_box_fill, height=46.0)
    _style_section_title_row(footprint_section_row, 2, 7, fill=secondary_section_fill, height=24.0)
    _style_section_title_row(footprint_section_row, 9, 19, fill=secondary_section_fill, height=24.0)
    _style_section_title_row(footprint_section_row, 21, 24, fill=diagnostic_section_fill, height=24.0)
    for box_start, box_end, _box_text in help_boxes:
        _style_box_range(
            box_start,
            box_end,
            21,
            24,
            fill=diagnostic_note_fill,
            font=note_font,
            alignment=note_alignment,
            border=thin_border,
        )

    _style_section_title_row(role_summary_title_row, 21, 24, fill=diagnostic_section_fill, height=24.0)
    _style_note_box_row(role_summary_note_row, 21, 24, fill=diagnostic_note_fill, height=36.0)
    _style_section_title_row(winner_story_title_row, 21, 24, fill=diagnostic_section_fill, height=24.0)
    _style_section_title_row(experimental_title_row, 21, 24, fill=diagnostic_section_fill, height=24.0)
    _style_section_title_row(offset_section_row, 2, 15, fill=diagnostic_section_fill, height=24.0)
    _style_section_title_row(table_start_row, 2, 21, fill=diagnostic_section_fill, height=24.0)

    if approx_market_crush_build_up_layout:
        build_title_row = int(approx_market_crush_build_up_layout.get("title_row") or 0)
        build_note_row = int(approx_market_crush_build_up_layout.get("note_row") or 0)
        build_header_row = int(approx_market_crush_build_up_layout.get("header_row") or 0)
        build_subheader_row = int(approx_market_crush_build_up_layout.get("subheader_row") or 0)
        build_end_row = int(approx_market_crush_build_up_layout.get("next_row") or 0)
        if build_title_row:
            _style_section_title_row(build_title_row, 2, 15, fill=primary_section_fill, height=24.0)
        if build_note_row:
            _style_note_box_row(build_note_row, 2, 15, fill=note_box_fill, height=28.0)
        if build_header_row:
            _style_row_range(
                build_header_row,
                2,
                15,
                fill=primary_header_fill,
                font=bold_font,
                alignment=align_center_wrap,
                border=thin_border,
                height=26.0,
            )
        if build_subheader_row:
            _style_row_range(
                build_subheader_row,
                2,
                15,
                fill=PatternFill("solid", fgColor="F8FBFD"),
                font=body_font,
                alignment=align_center_wrap,
                border=thin_border,
                height=24.0,
            )
            for cc in (2, 12):
                ws.cell(row=build_subheader_row, column=cc).alignment = copy(frame_text_alignment)
        econ_rows_local = approx_market_crush_build_up_layout.get("econ_rows") if isinstance(approx_market_crush_build_up_layout.get("econ_rows"), dict) else {}
        snapshot_focus_row = int(econ_rows_local.get("corn_basis_snapshot_date") or 0)
        selection_focus_row = int(econ_rows_local.get("corn_basis_selection_rule") or 0)
        for rr in (snapshot_focus_row, selection_focus_row):
            if rr:
                _style_row_range(
                    rr,
                    2,
                    15,
                    fill=focus_fill,
                    font=body_font,
                    alignment=align_center,
                    border=thin_border,
                    height=26.0,
                )
                ws.cell(row=rr, column=2).font = copy(bold_font)
                ws.cell(row=rr, column=2).alignment = copy(frame_text_alignment)
                ws.cell(row=rr, column=12).alignment = copy(note_alignment)
        if snapshot_focus_row:
            for cc in (3, 5, 7, 9, 11):
                ws.cell(row=snapshot_focus_row, column=cc).alignment = copy(frame_value_alignment)
        if selection_focus_row:
            for cc in (3, 5, 7, 9, 11):
                ws.cell(row=selection_focus_row, column=cc).alignment = copy(frame_text_alignment)
            ws.row_dimensions[selection_focus_row].height = 32.0
        if build_title_row and build_end_row:
            _apply_outer_border(build_title_row, build_end_row, 2, 15)

    if corn_oil_gate_check_layout:
        gate_title_row = int(corn_oil_gate_check_layout.get("title_row") or 0)
        gate_header_row = int(corn_oil_gate_check_layout.get("header_row") or 0)
        gate_note_row = int(corn_oil_gate_check_layout.get("note_row") or 0)
        provenance_title_row = int(corn_oil_gate_check_layout.get("provenance_title_row") or 0)
        provenance_header_row = int(corn_oil_gate_check_layout.get("provenance_header_row") or 0)
        gate_end_row = int(corn_oil_gate_check_layout.get("next_row") or 0)
        if gate_title_row:
            _style_section_title_row(gate_title_row, 2, 15, fill=primary_section_fill, height=24.0)
        if gate_header_row:
            _style_row_range(
                gate_header_row,
                2,
                15,
                fill=primary_header_fill,
                font=bold_font,
                alignment=align_center_wrap,
                border=thin_border,
                height=26.0,
            )
        if gate_note_row:
            _style_note_box_row(gate_note_row, 2, 15, fill=note_box_fill, height=36.0)
        if provenance_title_row:
            _style_section_title_row(provenance_title_row, 2, 15, fill=secondary_section_fill, height=23.0)
        if provenance_header_row:
            _style_row_range(
                provenance_header_row,
                2,
                15,
                fill=secondary_header_fill,
                font=bold_font,
                alignment=align_center_wrap,
                border=thin_border,
                height=26.0,
            )
        for rr in list((corn_oil_gate_check_layout.get("gate_rows") or {}).values()):
            for cc in (6,):
                ws.cell(row=int(rr), column=cc).alignment = copy(align_center_wrap)
            ws.row_dimensions[int(rr)].height = max(float(ws.row_dimensions[int(rr)].height or 0.0), 30.0)
        for rr in list((corn_oil_gate_check_layout.get("provenance_rows") or {}).values()):
            ws.cell(row=int(rr), column=6).alignment = copy(align_center_wrap)
            ws.row_dimensions[int(rr)].height = max(float(ws.row_dimensions[int(rr)].height or 0.0), 30.0)
        if gate_title_row and gate_end_row:
            _apply_outer_border(gate_title_row, gate_end_row, 2, 15)

    if coproduct_frame_summary_layout:
        frame_title_row = int(coproduct_frame_summary_layout.get("title_row") or 0)
        frame_note_row = int(coproduct_frame_summary_layout.get("note_row") or 0)
        frame_header_row = int(coproduct_frame_summary_layout.get("header_row") or 0)
        frame_end_row = int(coproduct_frame_summary_layout.get("next_row") or 0)
        if frame_title_row:
            _style_section_title_row(frame_title_row, 2, 10, fill=primary_section_fill, height=24.0)
        if frame_note_row:
            _style_note_box_row(frame_note_row, 2, 10, fill=note_box_fill, height=64.0)
        if frame_header_row:
            _style_row_range(
                frame_header_row,
                2,
                10,
                fill=primary_header_fill,
                font=bold_font,
                alignment=align_center_wrap,
                border=thin_border,
                height=38.0,
            )
        for rr in list((coproduct_frame_summary_layout.get("frame_rows") or {}).values()):
            rr_num = int(rr)
            ws.row_dimensions[rr_num].height = 36.0
            for cc in (3, 4, 5, 6, 7, 9):
                ws.cell(row=rr_num, column=cc).alignment = copy(numeric_alignment)
            for cc in (8, 10):
                ws.cell(row=rr_num, column=cc).alignment = copy(note_alignment)
        if frame_title_row and frame_end_row:
            _apply_outer_border(frame_title_row, frame_end_row, 2, 10)

    if coproduct_history_layout:
        history_title_row = int(coproduct_history_layout.get("title_row") or 0)
        history_note_row = int(coproduct_history_layout.get("note_row") or 0)
        history_header_row = int(coproduct_history_layout.get("header_row") or 0)
        first_history_row = int(coproduct_history_layout.get("first_data_row") or 0)
        last_history_row = int(coproduct_history_layout.get("last_data_row") or 0)
        if history_title_row:
            _style_section_title_row(history_title_row, 2, 9, fill=secondary_section_fill, height=24.0)
        if history_note_row:
            _style_note_box_row(history_note_row, 2, 9, fill=note_box_fill, height=56.0)
        if history_header_row:
            _style_row_range(
                history_header_row,
                2,
                9,
                fill=secondary_header_fill,
                font=bold_font,
                alignment=align_center_wrap,
                border=thin_border,
                height=32.0,
            )
        if first_history_row and last_history_row:
            for history_row_num in range(first_history_row, last_history_row + 1):
                row_fill = zebra_fill_light if ((history_row_num - history_header_row) % 2) else zebra_fill_dark
                _style_row_range(history_row_num, 2, 9, fill=row_fill, font=body_font, border=thin_border, height=26.0)
                ws.cell(row=history_row_num, column=2).alignment = copy(frame_text_alignment)
                for cc in (3, 4, 5, 6, 7, 9):
                    ws.cell(row=history_row_num, column=cc).alignment = copy(numeric_alignment)
                ws.cell(row=history_row_num, column=8).alignment = copy(note_alignment)
        if history_title_row and last_history_row:
            _apply_outer_border(history_title_row, last_history_row, 2, 9)

    if coproduct_volume_support_layout:
        volume_title_row = int(coproduct_volume_support_layout.get("title_row") or 0)
        volume_note_row = int(coproduct_volume_support_layout.get("note_row") or 0)
        volume_header_row = int(coproduct_volume_support_layout.get("header_row") or 0)
        volume_end_row = int(coproduct_volume_support_layout.get("next_row") or 0)
        if volume_title_row:
            _style_section_title_row(volume_title_row, 2, 15, fill=secondary_section_fill, height=24.0)
        if volume_note_row:
            _style_note_box_row(volume_note_row, 2, 15, fill=note_box_fill, height=34.0)
        if volume_header_row:
            _style_row_range(
                volume_header_row,
                2,
                15,
                fill=secondary_header_fill,
                font=bold_font,
                alignment=align_center_wrap,
                border=thin_border,
                height=30.0,
            )
        if volume_title_row and volume_end_row:
            _apply_outer_border(volume_title_row, volume_end_row, 2, 15)

    if coproduct_experimental_layout:
        exp_title_row = int(coproduct_experimental_layout.get("title_row") or 0)
        exp_note_row = int(coproduct_experimental_layout.get("note_row") or 0)
        exp_summary_header_row = int(coproduct_experimental_layout.get("summary_header_row") or 0)
        exp_comparison_header_row = int(coproduct_experimental_layout.get("comparison_header_row") or 0)
        exp_end_row = int(coproduct_experimental_layout.get("next_row") or 0)
        if exp_title_row:
            _style_section_title_row(exp_title_row, 2, 15, fill=diagnostic_section_fill, height=24.0)
        if exp_note_row:
            _style_note_box_row(exp_note_row, 2, 15, fill=diagnostic_note_fill, height=34.0)
        if exp_summary_header_row:
            _style_row_range(
                exp_summary_header_row,
                2,
                15,
                fill=secondary_header_fill,
                font=bold_font,
                alignment=align_center_wrap,
                border=thin_border,
                height=24.0,
            )
        if exp_comparison_header_row:
            _style_row_range(
                exp_comparison_header_row,
                2,
                12,
                fill=primary_header_fill,
                font=bold_font,
                alignment=align_center_wrap,
                border=thin_border,
                height=36.0,
            )
        for rr in list((coproduct_experimental_layout.get("summary_rows_by_label") or {}).values()):
            rr_num = int(rr)
            ws.row_dimensions[rr_num].height = max(float(ws.row_dimensions[rr_num].height or 0.0), 28.0)
            ws.cell(row=rr_num, column=2).font = copy(bold_font)
            ws.cell(row=rr_num, column=2).alignment = copy(frame_text_alignment)
            ws.cell(row=rr_num, column=3).alignment = copy(note_alignment)
        for rr in list((coproduct_experimental_layout.get("comparison_rows_by_key") or {}).values()):
            rr_num = int(rr)
            ws.row_dimensions[rr_num].height = 34.0
            for cc in (4, 5, 6, 10, 11):
                ws.cell(row=rr_num, column=cc).alignment = copy(numeric_alignment)
            ws.cell(row=rr_num, column=7).alignment = copy(frame_value_alignment)
            ws.cell(row=rr_num, column=8).alignment = copy(frame_text_alignment)
            ws.cell(row=rr_num, column=9).alignment = copy(frame_text_alignment)
            ws.cell(row=rr_num, column=12).alignment = copy(align_center_wrap)
        if exp_title_row and exp_end_row:
            _apply_outer_border(exp_title_row, exp_end_row, 2, 15)

    _style_section_title_row(memo_section_row, 2, 15, fill=diagnostic_section_fill, height=24.0)
    _style_section_title_row(hedge_section_row, 2, 15, fill=diagnostic_section_fill, height=24.0)
    _style_note_box_row(hedge_summary_row, 2, 15, fill=diagnostic_note_fill, height=34.0)
    _style_note_box_row(hedge_note_row, 2, 15, fill=diagnostic_note_fill, height=36.0)
    _style_section_title_row(hedge_leader_title_row, 2, 8, fill=diagnostic_section_fill, height=23.0)
    _style_section_title_row(hedge_quarter_title_row, 2, 9, fill=diagnostic_section_fill, height=23.0)
    _style_section_title_row(hedge_interp_title_row, 2, 15, fill=diagnostic_section_fill, height=23.0)
    _style_note_box_row(hedge_interp_row, 2, 15, fill=diagnostic_note_fill, height=46.0)
    _apply_outer_border(memo_section_row, hedge_interp_row, 2, 15)
    _style_section_title_row(futures_section_row, 2, 16, fill=diagnostic_section_fill, height=24.0)
    _style_note_box_row(futures_note_row, 2, 16, fill=diagnostic_note_fill, height=42.0)
    _style_section_title_row(futures_summary_title_row, 2, 12, fill=diagnostic_section_fill, height=23.0)
    _style_section_title_row(futures_detail_title_row, 2, 16, fill=diagnostic_section_fill, height=23.0)
    _apply_outer_border(futures_section_row, futures_detail_end_row, 2, 16)

    for spacer_row in {
        int((approx_market_crush_build_up_layout or {}).get("title_row") or 0) - 1,
        int((corn_oil_gate_check_layout or {}).get("title_row") or 0) - 1,
        int((coproduct_frame_summary_layout or {}).get("title_row") or 0) - 1,
        int((coproduct_history_layout or {}).get("title_row") or 0) - 1,
        int((coproduct_volume_support_layout or {}).get("title_row") or 0) - 1,
        int((coproduct_experimental_layout or {}).get("title_row") or 0) - 1,
        memo_section_row - 1,
        hedge_section_row - 1,
        hedge_leader_title_row - 1,
        hedge_quarter_title_row - 1,
        hedge_interp_title_row - 1,
        futures_section_row - 1,
        futures_summary_title_row - 1,
        futures_detail_title_row - 1,
    }:
        if spacer_row <= 0:
            continue
        _style_row_range(spacer_row, 2, 24, fill=spacer_fill, border=Border(), height=12.0)
    return {
        "approx_market_crush_build_up": approx_market_crush_build_up_layout,
        "corn_oil_gate_check": corn_oil_gate_check_layout,
        "coproduct_frame_summary": coproduct_frame_summary_layout,
        "coproduct_signal_readiness": coproduct_signal_readiness_layout,
        "coproduct_quarterly_history": coproduct_history_layout,
        "coproduct_volume_support_audit": coproduct_volume_support_layout,
        "coproduct_experimental_lenses": coproduct_experimental_layout,
        "futures_timing_sandbox": {
            "title_row": futures_section_row,
            "summary_header_row": futures_summary_header_row,
            "detail_header_row": futures_detail_header_row,
            "next_row": futures_detail_end_row,
        },
        "coproduct_visible_block_allowed": bool((corn_oil_gate_check_layout or {}).get("overlay_activation_pass")),
    }
