"""BS_Segments sheet writer extracted from excel_writer_context."""
from __future__ import annotations

import html
import math
import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Callable, Dict, List, Mapping, Optional, Pattern, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .cache_layout import ticker_cache_roots_from_base_dir
from .excel_writer_coloring import (
    _hidden_source_comparison_metric,
    _quarterly_color_metric_from_series,
    _quarterly_row_color_policy,
)
from .excel_writer_segment_sources import (
    _annual_segment_latest_year_for_qa,
    _anf_add_total_company_quarter_revenue_from_history,
    _anf_annual_segment_data_from_slides_segments,
    _anf_fill_brand_quarter_revenue_from_annual_segments_for_bs,
    _filter_anf_quarterly_segment_actual_rows,
    _pbi_add_corporate_reconciliation_from_release_text,
    _pbi_repair_total_reportable_segment_quarterly_totals_for_bs,
)
from .excel_writer_segments import (
    annual_segment_label as ew_annual_segment_label,
    extract_quarter_from_cell as ew_extract_quarter_from_cell,
    extract_segment_line_values as ew_extract_segment_line_values,
    extract_year_from_cell as ew_extract_year_from_cell,
    latest_segment_financials_workbook as ew_latest_segment_financials_workbook,
    parse_annual_segment_data_from_workbook as ew_parse_annual_segment_data_from_workbook,
    parse_quarterly_segment_data_from_workbook as ew_parse_quarterly_segment_data_from_workbook,
    quarterly_segment_label as ew_quarterly_segment_label,
)
from .excel_writer_sources import infer_q_from_name as source_infer_q_from_name
from .guidance_lexicon import normalize_text as glx_normalize_text
from .legacy_support import (
    _extract_balance_sheet_from_html,
    _extract_balance_sheet_from_text,
    _path_belongs_to_ticker,
)
from .non_gaap import infer_quarter_end_from_text, strip_html


@dataclass
class BsSegmentsWriterDeps:
    wb: Workbook
    hist: Optional[pd.DataFrame]
    audit: Optional[pd.DataFrame]
    ticker: Any
    company_profile: Any
    slides_segments: Optional[pd.DataFrame]
    material_roots: Sequence[Path]
    ticker_roots: Sequence[Path]
    ui_info_rows: List[Dict[str, Any]]
    font_size: float
    header_size: float
    is_pbi_profile: bool
    is_gpre_profile: bool
    is_anf_profile: bool
    bank_metrics_enabled: bool
    enable_quarterly_segment_block: bool
    enable_annual_segment_block: bool
    quarterly_segment_labels: Sequence[str]
    annual_segment_labels: Sequence[str]
    annual_segment_alias_patterns: Sequence[Tuple[Pattern[str], str]]
    anf_segment_brand_explanation: str
    get_valuation_style_bundle: Callable[[], Dict[str, Any]]
    hist_view: Callable[..., pd.DataFrame]
    resolve_col: Callable[[pd.DataFrame, List[str]], Optional[str]]
    set_cell_comment: Callable[..., None]
    shared_load_local_balance_sheet_detail_payloads: Callable[..., Dict[date, Dict[str, Any]]]
    carry_forward_low_change_series: Callable[..., Dict[pd.Timestamp, Optional[float]]]
    first_existing_material_dir: Callable[..., Optional[Path]]
    parse_quarter_from_filename: Callable[[str], Optional[date]]
    parse_quarter_from_follow_text: Callable[[str], Optional[date]]
    read_operating_driver_text: Callable[[Path], str]
    operating_driver_financial_statement_files: Callable[[], List[Path]]
    sec_cache_roots_local: Callable[[], List[Path]]
    anf_visible_quarter_label: Callable[[Any], str]


def _series_has_nonblank_values(values: Mapping[Any, Any] | None) -> bool:
    """Return True when a source/value map has at least one intentional value."""
    for raw_val in dict(values or {}).values():
        if raw_val is None:
            continue
        if isinstance(raw_val, str) and not raw_val.strip():
            continue
        try:
            if pd.isna(raw_val):
                continue
        except (TypeError, ValueError):
            pass
        return True
    return False


def _should_render_carbon_equipment_liabilities(ticker: Any, values: Mapping[Any, Any] | None) -> bool:
    """Carbon equipment liabilities are sector-specific unless source data exists."""
    ticker_txt = str(ticker or "").strip().upper()
    return ticker_txt == "GPRE" or _series_has_nonblank_values(values)



def write_bs_segments_sheet(deps: BsSegmentsWriterDeps, quarters_shown: int = 8) -> List[Dict[str, Any]]:

    wb = deps.wb
    hist = deps.hist
    audit = deps.audit
    ticker = deps.ticker
    company_profile = deps.company_profile
    slides_segments = deps.slides_segments
    material_roots = deps.material_roots
    ticker_roots = deps.ticker_roots
    ui_info_rows = deps.ui_info_rows
    font_size = deps.font_size
    header_size = deps.header_size
    is_pbi_profile = deps.is_pbi_profile
    is_gpre_profile = deps.is_gpre_profile
    is_anf_profile = deps.is_anf_profile
    bank_metrics_enabled = deps.bank_metrics_enabled
    enable_quarterly_segment_block = deps.enable_quarterly_segment_block
    enable_annual_segment_block = deps.enable_annual_segment_block
    quarterly_segment_labels = deps.quarterly_segment_labels
    annual_segment_labels = deps.annual_segment_labels
    annual_segment_alias_patterns = deps.annual_segment_alias_patterns
    ANF_SEGMENT_BRAND_EXPLANATION = deps.anf_segment_brand_explanation
    _get_valuation_style_bundle = deps.get_valuation_style_bundle
    _hist_view = deps.hist_view
    _resolve_col = deps.resolve_col
    _set_cell_comment_local = deps.set_cell_comment
    _shared_load_local_balance_sheet_detail_payloads = deps.shared_load_local_balance_sheet_detail_payloads
    _carry_forward_low_change_series = deps.carry_forward_low_change_series
    _first_existing_material_dir = deps.first_existing_material_dir
    _parse_quarter_from_filename = deps.parse_quarter_from_filename
    _parse_quarter_from_follow_text = deps.parse_quarter_from_follow_text
    _read_operating_driver_text = deps.read_operating_driver_text
    _operating_driver_financial_statement_files = deps.operating_driver_financial_statement_files
    _sec_cache_roots_local = deps.sec_cache_roots_local
    _anf_visible_quarter_label = deps.anf_visible_quarter_label
    qa_rows: List[Dict[str, Any]] = []
    if "BS_Segments" in wb.sheetnames:
        old = wb["BS_Segments"]
        idx_old = wb._sheets.index(old)
        wb.remove(old)
        ws = wb.create_sheet("BS_Segments", idx_old)
    else:
        ws = wb.create_sheet("BS_Segments")

    ws.sheet_format.defaultRowHeight = 18
    ws.sheet_view.zoomScale = 110

    ws_val = wb["Valuation"] if "Valuation" in wb.sheetnames else None
    style_bundle = _get_valuation_style_bundle()
    header_fill = copy(style_bundle["header_fill"])
    section_fill = copy(style_bundle["section_fill"])
    bold = copy(style_bundle["bold_font"])
    title_fill = copy(style_bundle["title_fill"])
    thin_border = copy(style_bundle["thin_border"])
    valuation_quarter_style_a = style_bundle.get("valuation_quarter_style_a")
    valuation_quarter_style_col = style_bundle.get("valuation_quarter_style_col")
    valuation_actuals_style_col = style_bundle.get("valuation_actuals_style_col")
    valuation_section_label_style = style_bundle.get("valuation_section_label_style")
    valuation_section_col_style = style_bundle.get("valuation_section_col_style")
    valuation_label_style = style_bundle.get("valuation_label_style")
    valuation_numeric_style = style_bundle.get("valuation_numeric_style")
    valuation_bucket_fills: Dict[str, PatternFill] = dict(style_bundle.get("valuation_bucket_fills") or {})
    valuation_col_widths: Dict[str, Optional[float]] = dict(style_bundle.get("valuation_col_widths") or {})
    valuation_row_height_actual: Optional[float] = style_bundle.get("valuation_row_height_actual")
    valuation_row_height_quarter: Optional[float] = style_bundle.get("valuation_row_height_quarter")
    valuation_data_font_size: float = float(style_bundle.get("valuation_data_font_size") or font_size)

    if ws_val is not None:
        for cc in range(1, 8):
            src = ws_val.cell(row=1, column=cc)
            dst = ws.cell(row=1, column=cc)
            dst.value = src.value
            dst._style = copy(src._style)
    else:
        ws["A1"] = "Scale"
        ws["B1"] = "$m"
    ws["A2"] = "Values scaled to $m unless %"
    ws["A4"] = "QA: pending"
    ws["A4"].font = Font(bold=True, size=11)

    if hist is None or hist.empty or "quarter" not in hist.columns:
        ws["A4"] = "No data."
        return qa_rows

    h = _hist_view().copy()
    if "_quarter" in h.columns:
        h["quarter"] = h["_quarter"]
    h = h[h["quarter"].notna()].sort_values("quarter")
    if h.empty:
        ws["A4"] = "No data."
        return qa_rows

    qs = sorted(h["quarter"].dropna().unique())
    qs = qs[-int(max(1, quarters_shown)) :]
    start_col = 2
    actuals_row = 10
    quarter_row = 11
    row_idx = 12
    last_col = start_col + len(qs) - 1
    last_col_letter = get_column_letter(last_col)

    title_end_col = max(last_col, 9)
    if is_anf_profile:
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=title_end_col)
        ws["A5"] = "ANF quarter labels are fiscal periods; Q4 2025 ended 2026-01-31."
        ws["A5"].font = Font(italic=True, size=10, color="666666")
        ws["A5"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        ws.row_dimensions[5].height = 16.0
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=title_end_col)
    ws["A8"] = "Balance sheet & Segments"
    ws["A8"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A8"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A8"].fill = title_fill
    ws.row_dimensions[8].height = 24

    ws.merge_cells(start_row=actuals_row, start_column=start_col, end_row=actuals_row, end_column=last_col)
    ws[f"{get_column_letter(start_col)}{actuals_row}"] = "Actuals"
    ws[f"{get_column_letter(start_col)}{actuals_row}"].font = bold
    ws[f"{get_column_letter(start_col)}{actuals_row}"].alignment = Alignment(horizontal="center")
    ws[f"A{quarter_row}"] = "Quarter"
    if valuation_quarter_style_a is not None:
        ws[f"A{quarter_row}"]._style = copy(valuation_quarter_style_a)
    else:
        ws[f"A{quarter_row}"].font = bold
        ws[f"A{quarter_row}"].fill = header_fill
    for i, q in enumerate(qs):
        col = start_col + i
        col_letter = get_column_letter(col)
        qd = pd.Timestamp(q).date()
        ws[f"{col_letter}{quarter_row}"] = _anf_visible_quarter_label(qd) if is_anf_profile else f"{qd.year}-Q{((qd.month - 1) // 3) + 1}"
        if valuation_quarter_style_col is not None:
            ws[f"{col_letter}{quarter_row}"]._style = copy(valuation_quarter_style_col)
        else:
            ws[f"{col_letter}{quarter_row}"].font = bold
            ws[f"{col_letter}{quarter_row}"].alignment = Alignment(horizontal="center")
            ws[f"{col_letter}{quarter_row}"].fill = header_fill
            ws[f"{col_letter}{quarter_row}"].border = thin_border
        if valuation_actuals_style_col is not None:
            ws[f"{col_letter}{actuals_row}"]._style = copy(valuation_actuals_style_col)
        else:
            ws[f"{col_letter}{actuals_row}"].fill = header_fill
            ws[f"{col_letter}{actuals_row}"].border = thin_border
    if valuation_row_height_actual is not None:
        ws.row_dimensions[actuals_row].height = valuation_row_height_actual
    if valuation_row_height_quarter is not None:
        ws.row_dimensions[quarter_row].height = valuation_row_height_quarter
    ws.row_dimensions[actuals_row].height = 19.5
    ws.row_dimensions[quarter_row].height = 19.5

    h_idx = h.set_index("quarter")

    def _load_local_balance_sheet_detail_payloads(target_quarters: set[date]) -> Dict[date, Dict[str, Any]]:
        out_local: Dict[date, Dict[str, Any]] = {}
        for path_in in _operating_driver_financial_statement_files():
            raw_txt = _read_operating_driver_text(path_in)
            if not raw_txt:
                continue
            qd = _parse_quarter_from_filename(path_in.name) or _parse_quarter_from_follow_text(raw_txt) or infer_quarter_end_from_text(raw_txt)
            if not isinstance(qd, date) or qd not in target_quarters:
                continue
            result = None
            try:
                if path_in.suffix.lower() in {".htm", ".html"}:
                    result = _extract_balance_sheet_from_html(path_in.read_bytes(), qd)
                else:
                    result = _extract_balance_sheet_from_text(raw_txt, qd)
            except Exception:
                result = None
            if not result:
                continue
            payload = dict(result)
            payload["source_doc"] = str(path_in)
            if qd not in out_local or len(payload.get("values", {})) >= len(out_local[qd].get("values", {})):
                out_local[qd] = payload
        return out_local

    def _extract_segment_line_values(line_in: Any, year_count: int, *, exact_count: bool = False) -> List[float]:
        return ew_extract_segment_line_values(line_in, year_count, exact_count=exact_count)

    def _annual_segment_label(metric_label: str, line_in: Any) -> str:
        return ew_annual_segment_label(
            metric_label,
            line_in,
            annual_segment_alias_patterns=annual_segment_alias_patterns,
            company_segment_alias_patterns=company_profile.segment_alias_patterns,
        )

    def _extract_year_from_cell(value_in: Any) -> Optional[int]:
        return ew_extract_year_from_cell(value_in)

    def _latest_segment_financials_workbook() -> Optional[Path]:
        return ew_latest_segment_financials_workbook(
            _first_existing_material_dir("segment_financials", "historical_segment")
        )

    def _extract_quarter_from_cell(value_in: Any) -> Optional[pd.Timestamp]:
        return ew_extract_quarter_from_cell(value_in)

    def _quarterly_segment_label(metric_label: str, line_in: Any) -> str:
        return ew_quarterly_segment_label(
            metric_label,
            line_in,
            annual_segment_alias_patterns=annual_segment_alias_patterns,
            company_segment_alias_patterns=company_profile.segment_alias_patterns,
        )

    def _parse_quarterly_segment_data_from_workbook(path_in: Path) -> Dict[str, Any]:
        return ew_parse_quarterly_segment_data_from_workbook(
            path_in,
            annual_segment_alias_patterns=annual_segment_alias_patterns,
            company_segment_alias_patterns=company_profile.segment_alias_patterns,
        )

    def _parse_annual_segment_data_from_workbook(path_in: Path) -> Dict[str, Any]:
        return ew_parse_annual_segment_data_from_workbook(
            path_in,
            annual_segment_alias_patterns=annual_segment_alias_patterns,
            company_segment_alias_patterns=company_profile.segment_alias_patterns,
        )

    def _parse_annual_segment_data_from_text(text_in: Any) -> Dict[str, Any]:
        raw_txt = html.unescape(str(text_in or ""))
        if not raw_txt:
            return {}
        if re.search(r"<[^>]+>", raw_txt):
            try:
                raw_txt = strip_html(raw_txt)
            except Exception:
                raw_txt = re.sub(r"<[^>]+>", " ", raw_txt)
        out: Dict[str, Any] = {"metrics": {}, "assets": {}, "years": []}
        allowed_annual_segment_labels = {
            str(lbl).strip(): str(lbl).strip()
            for lbl in (annual_segment_labels or tuple())
            if str(lbl or "").strip()
        }
        allowed_annual_segment_labels_lc = {
            label.lower(): label for label in allowed_annual_segment_labels.values()
        }
        revenue_detail_rows = {
            "revenues from external customers",
            "intersegment revenues",
            "total segment revenues",
            "revenues including intersegment activity",
        }
        annual_segment_noise_re = re.compile(
            r"\b(ebitda margin|adjusted segment ebitda|pbi adj(?:u)?ted ebitda|"
            r"adjusted segment ebit|pbi adjusted ebit)\b",
            re.I,
        )
        raw_txt = raw_txt.replace("\r", "")
        raw_low = raw_txt.lower()
        if "selected operating segment financial information are as follows" in raw_low:
            seg_start = raw_low.find("selected operating segment financial information are as follows")
            assets_start = raw_low.find("total assets by segment are as follows", seg_start)
            seg_end = len(raw_txt)
            for marker in (
                "we use ebitda",
                "the following table reconciles net loss",
                "ethanol production includes inventory lower of cost or net realizable value adjustments",
            ):
                idx = raw_low.find(marker, seg_start)
                if idx >= 0 and (assets_start < 0 or idx < assets_start):
                    seg_end = min(seg_end, idx)
            segment_blocks = [raw_txt[seg_start:seg_end]]
            if assets_start >= 0:
                assets_end = len(raw_txt)
                assets_marker = "asset balances by segment exclude intercompany balances"
                assets_note_idx = raw_low.find(assets_marker, assets_start)
                if assets_note_idx >= 0:
                    assets_end = min(len(raw_txt), assets_note_idx + len(assets_marker) + 1)
                segment_blocks.append(raw_txt[assets_start:assets_end])
            raw_txt = "\n".join(block for block in segment_blocks if str(block or "").strip())
            raw_low = raw_txt.lower()
        if raw_txt.count("\n") < 20 and "selected operating segment financial information are as follows" in raw_low:
            for cue in [
                "The selected operating segment financial information are as follows",
                "Total assets by segment are as follows",
                "Cost of goods sold",
                "Gross margin",
                "Depreciation and amortization",
                "Operating income (loss)",
                "Ethanol production",
                "Agribusiness and energy services",
                "Corporate activities",
                "Corporate assets",
                "Revenues from external customers",
                "Intersegment revenues",
                "Total segment revenues",
                "Revenues including intersegment activity",
                "Intersegment eliminations",
            ]:
                raw_txt = re.sub(rf"\s*{re.escape(cue)}\s*", f"\n{cue}\n", raw_txt, flags=re.I)
            raw_txt = re.sub(
                r"(Year Ended December 31,)\s*((?:20\d{2}\s+){1,3}20\d{2})",
                lambda m: f"{m.group(1)}\n{m.group(2).strip()}\n",
                raw_txt,
                flags=re.I,
            )
            raw_txt = re.sub(
                r"((?:20\d{2}\s+){1,3}20\d{2})\s+"
                r"(Revenues|Cost of goods sold|Gross margin|Depreciation and amortization|Operating income \(loss\)|Total assets)\b",
                lambda m: f"{m.group(1).strip()}\n{m.group(2)}\n",
                raw_txt,
                flags=re.I,
            )
            raw_txt = re.sub(r"\n{2,}", "\n", raw_txt)
        lines = []
        for raw_line in raw_txt.splitlines():
            line = str(raw_line or "").replace("â€”", "—")
            line = re.sub(r"\(\d+\)", "", line)
            line = re.sub(r"\s+", " ", line).strip()
            if line:
                lines.append(line)
        if not lines:
            return {}

        def _line_years(line_in: str) -> List[int]:
            years = [int(y) for y in re.findall(r"\b20\d{2}\b", line_in)]
            if len(years) >= 2 and years == sorted(years, reverse=True):
                return years
            return []

        def _line_values(line_in: str, year_count: int, *, prefer_first: bool = False) -> List[float]:
            clean_line = str(line_in or "")
            for marker in (
                "Year Ended December 31,",
                "Asset balances by segment exclude intercompany balances",
            ):
                marker_idx = clean_line.lower().find(marker.lower())
                if marker_idx > 0:
                    clean_line = clean_line[:marker_idx].strip()
            clean_line = re.sub(r"\b\d+\s+T\s*a\s*b\s+le\s+of\s+Contents\b.*$", "", clean_line, flags=re.I)
            vals: List[float] = []
            for mm in re.finditer(
                r"(?<!\d)(?:\(?-?[0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?\)?)(?!\d)",
                clean_line,
            ):
                token = str(mm.group(0) or "").replace("(", "-").replace(")", "").replace(",", "")
                try:
                    vals.append(float(token) * 1000.0)
                except Exception:
                    continue
            if len(vals) < year_count:
                return []
            return list(vals[:year_count] if prefer_first else vals[-year_count:])

        def _segment_label_for_line(metric_label: str, line_in: str) -> str:
            seg_label = _annual_segment_label(metric_label, line_in)
            if not seg_label:
                return ""
            seg_label = str(seg_label).strip()
            if not seg_label:
                return ""
            raw_line = str(line_in or "").strip()
            raw_low = raw_line.lower()
            seg_low = seg_label.lower()
            if seg_low in revenue_detail_rows or seg_low in {"cost of goods sold", "total assets"}:
                return ""
            canonical_label = allowed_annual_segment_labels_lc.get(seg_low, "")
            if canonical_label:
                return canonical_label
            if seg_label == raw_line and re.search(r"\d", raw_line):
                return ""
            if raw_low == seg_low:
                return ""
            return ""

        metric_map = {
            "revenues": "Revenues",
            "gross_margin": "Gross margin",
            "depreciation_amortization": "Depreciation & amortization",
            "operating_income_loss": "Operating income (loss)",
        }
        in_financials = False
        in_assets = False
        section_key = ""
        section_years: List[int] = []
        asset_years: List[int] = []
        current_segment = ""
        current_asset_segment = ""
        for line in lines:
            low = line.lower()
            if annual_segment_noise_re.search(line):
                continue
            if "selected operating segment financial information are as follows" in low:
                in_financials = True
                in_assets = False
                section_key = ""
                section_years = []
                current_segment = ""
                continue
            if "total assets by segment are as follows" in low:
                in_financials = False
                in_assets = True
                asset_years = []
                current_asset_segment = ""
                continue
            if low.startswith("we use ebitda") or low.startswith("the following table reconciles net loss"):
                in_financials = False
                current_segment = ""
            if re.match(r"^\(\d+\)\s", line):
                in_financials = False
                current_segment = ""
                continue
            if low.startswith("year ended december 31, ") and "compared with" in low:
                in_assets = False
                current_asset_segment = ""
            if low.startswith("the following discussion provides greater detail") or low.startswith("key operating data for our ethanol production segment is as follows"):
                in_assets = False
                current_asset_segment = ""
            if line.startswith("F-") or low.startswith("table of contents"):
                continue

            if in_financials:
                years = _line_years(line)
                if low.startswith("year ended december 31"):
                    continue
                if years:
                    section_years = years
                    out["years"] = sorted(set(list(out["years"]) + years))
                    continue
                if low == "revenues":
                    section_key = "revenues"
                    current_segment = ""
                    continue
                if low == "cost of goods sold":
                    section_key = ""
                    current_segment = ""
                    continue
                if low == "gross margin":
                    section_key = "gross_margin"
                    current_segment = ""
                    continue
                if low == "depreciation and amortization":
                    section_key = "depreciation_amortization"
                    current_segment = ""
                    continue
                if low.startswith("operating income"):
                    section_key = "operating_income_loss"
                    current_segment = ""
                    continue
                if line.startswith("$") and not current_segment:
                    continue
                if low in revenue_detail_rows:
                    continue
                vals = _line_values(line, len(section_years), prefer_first=True) if section_years else []
                seg = _segment_label_for_line(
                    "Revenues" if section_key == "revenues" else metric_map.get(section_key, ""),
                    line,
                )
                if section_key == "revenues":
                    if seg == "Intersegment eliminations" and len(vals) == len(section_years):
                        for yy, vv in zip(section_years, vals):
                            out["metrics"].setdefault("Revenues", {}).setdefault(seg, {})[yy] = vv
                        continue
                    if seg == "Intersegment eliminations":
                        current_segment = seg
                        continue
                    if seg and seg != "Intersegment eliminations":
                        current_segment = seg
                        continue
                    if current_segment == "Intersegment eliminations" and len(vals) == len(section_years):
                        for yy, vv in zip(section_years, vals):
                            out["metrics"].setdefault("Revenues", {}).setdefault(current_segment, {})[yy] = vv
                        current_segment = ""
                    continue
                elif section_key in {"gross_margin", "depreciation_amortization", "operating_income_loss"} and seg and section_years:
                    if any(
                        cue in low
                        for cue in [
                            " includes ",
                            " include ",
                            " compared with ",
                            " as a result ",
                            " primarily due ",
                            " primarily as a result ",
                            " for the year ended ",
                        ]
                    ):
                        current_segment = ""
                        continue
                    if len(line) > 120 and not vals:
                        current_segment = ""
                        continue
                    current_segment = seg
                    if len(vals) != len(section_years):
                        continue
                    metric_label = metric_map.get(section_key, section_key)
                    for yy, vv in zip(section_years, vals):
                        out["metrics"].setdefault(metric_label, {}).setdefault(current_segment, {})[yy] = vv
                    current_segment = ""
                elif section_key in {"gross_margin", "depreciation_amortization", "operating_income_loss"} and current_segment and section_years:
                    if len(vals) != len(section_years):
                        continue
                    metric_label = metric_map.get(section_key, section_key)
                    for yy, vv in zip(section_years, vals):
                        out["metrics"].setdefault(metric_label, {}).setdefault(current_segment, {})[yy] = vv
                    current_segment = ""

            if in_assets:
                years = _line_years(line)
                if low.startswith("year ended december 31"):
                    continue
                if years:
                    asset_years = years
                    out["years"] = sorted(set(list(out["years"]) + years))
                    continue
                if low.startswith("total assets") or (line.startswith("$") and not current_asset_segment):
                    continue
                seg = _segment_label_for_line("Total assets", line)
                vals = _line_values(line, len(asset_years), prefer_first=True) if asset_years else []
                if seg and asset_years:
                    current_asset_segment = seg
                    if len(vals) != len(asset_years):
                        continue
                    for yy, vv in zip(asset_years, vals):
                        out["assets"].setdefault(current_asset_segment, {})[yy] = vv
                    current_asset_segment = ""
                elif current_asset_segment and asset_years:
                    if len(vals) != len(asset_years):
                        continue
                    for yy, vv in zip(asset_years, vals):
                        out["assets"].setdefault(current_asset_segment, {})[yy] = vv
                    current_asset_segment = ""
        return out

    def _annual_segment_candidate_quarter(path_in: Path, raw_txt: Any = "") -> Optional[date]:
        qd = (
            _parse_quarter_from_filename(path_in.name)
            or _parse_quarter_from_follow_text(raw_txt)
            or infer_quarter_end_from_text(raw_txt)
        )
        if isinstance(qd, date):
            return qd
        nm = path_in.name.lower()
        compact_year_end_match = re.search(r"(20\d{2})[-_]?12[-_]?31", nm)
        if compact_year_end_match:
            try:
                return date(int(compact_year_end_match.group(1)), 12, 31)
            except Exception:
                return None
        annual_match = re.search(r"(20\d{2})", nm)
        if annual_match and any(tok in nm for tok in ("annual_report", "annual report", "annualreport")):
            try:
                return date(int(annual_match.group(1)), 12, 31)
            except Exception:
                return None
        return None

    def _annual_segment_text_source_files() -> List[Path]:
        files: List[Path] = []
        seen: set[str] = set()

        def _add_path(path_in: Path) -> None:
            if path_in.suffix.lower() not in {".pdf", ".txt", ".htm", ".html"}:
                return
            if not _path_belongs_to_ticker(path_in, ticker, ticker_roots):
                return
            try:
                key = str(path_in.resolve())
            except Exception:
                key = str(path_in)
            if key in seen:
                return
            seen.add(key)
            files.append(path_in)

        for path_in in _operating_driver_financial_statement_files():
            _add_path(path_in)
        for root in material_roots:
            annual_dir = root / "annual_reports"
            if not annual_dir.exists() or not annual_dir.is_dir():
                continue
            try:
                cand_files = sorted([p for p in annual_dir.iterdir() if p.is_file()])
            except Exception:
                continue
            for path_in in cand_files:
                _add_path(path_in)

        sec_roots: List[Path] = []
        seen_roots: set[str] = set()

        def _add_sec_root(path_in: Path) -> None:
            if not path_in.exists() or not path_in.is_dir():
                return
            try:
                key = str(path_in.resolve())
            except Exception:
                key = str(path_in)
            if key in seen_roots:
                return
            seen_roots.add(key)
            sec_roots.append(path_in)

        for path_in in _sec_cache_roots_local():
            _add_sec_root(path_in)
        for root in material_roots:
            for sec_cache_dir in ticker_cache_roots_from_base_dir(root):
                _add_sec_root(sec_cache_dir)

        for sec_root in sec_roots:
            for pattern in (
                "*1231*.htm",
                "*1231*.html",
                "*1231*.txt",
                "*10-k*.htm",
                "*10-k*.html",
                "*10-k*.txt",
                "*10k*.htm",
                "*10k*.html",
                "*10k*.txt",
            ):
                try:
                    cand_files = sorted(sec_root.rglob(pattern))
                except Exception:
                    continue
                for path_in in cand_files:
                    if not path_in.is_file():
                        continue
                    _add_path(path_in)
        return files

    def _load_latest_annual_segment_data() -> Dict[str, Any]:
        if not enable_annual_segment_block:
            return {}
        workbook_path = _latest_segment_financials_workbook()
        if workbook_path is not None:
            parsed_wb = _parse_annual_segment_data_from_workbook(workbook_path)
            if parsed_wb.get("metrics") or parsed_wb.get("assets"):
                parsed_wb["source_doc"] = str(workbook_path)
                parsed_wb["source_qd"] = _parse_quarter_from_filename(workbook_path.name)
                return parsed_wb
        best_path: Optional[Path] = None
        best_qd: Optional[date] = None
        best_parsed: Optional[Dict[str, Any]] = None
        parsed_sources: List[Tuple[date, Path, Dict[str, Any]]] = []
        for path_in in _annual_segment_text_source_files():
            raw_txt = _read_operating_driver_text(path_in)
            if not raw_txt:
                continue
            raw_low = raw_txt.lower()
            if (
                "selected operating segment financial information are as follows" not in raw_low
                and "total assets by segment are as follows" not in raw_low
            ):
                continue
            parsed_preview = _parse_annual_segment_data_from_text(raw_txt)
            if not parsed_preview.get("metrics") and not parsed_preview.get("assets"):
                continue
            qd = _annual_segment_candidate_quarter(path_in, raw_txt)
            if not isinstance(qd, date):
                preview_years = [int(y) for y in parsed_preview.get("years") or [] if str(y).isdigit()]
                if preview_years:
                    qd = date(max(preview_years), 12, 31)
            if not isinstance(qd, date) or qd.month != 12:
                continue
            parsed_sources.append((qd, path_in, dict(parsed_preview)))
            if best_qd is None or qd > best_qd:
                best_qd = qd
                best_path = path_in
                best_parsed = dict(parsed_preview)
        if best_path is None or best_parsed is None:
            if is_anf_profile:
                return _anf_annual_segment_data_from_slides_segments(slides_segments)
            return {}
        parsed = dict(best_parsed)
        if is_gpre_profile:
            backfill_source_docs: List[str] = []
            parsed_assets: Dict[str, Dict[int, float]] = {
                str(seg): dict(year_vals or {})
                for seg, year_vals in dict(parsed.get("assets") or {}).items()
                if str(seg or "").strip()
            }
            parsed_years = {int(y) for y in parsed.get("years") or [] if str(y).isdigit()}
            for _qd_src, path_src, parsed_src in sorted(parsed_sources, key=lambda item: item[0], reverse=True):
                for seg, year_vals in dict(parsed_src.get("assets") or {}).items():
                    seg_key = str(seg or "").strip()
                    if not seg_key:
                        continue
                    target_vals = parsed_assets.setdefault(seg_key, {})
                    for yy, vv in dict(year_vals or {}).items():
                        try:
                            yy_int = int(yy)
                        except Exception:
                            continue
                        if yy_int in target_vals:
                            continue
                        vv_num = pd.to_numeric(vv, errors="coerce")
                        if pd.isna(vv_num):
                            continue
                        target_vals[yy_int] = float(vv_num)
                        parsed_years.add(yy_int)
                        src_doc = str(path_src)
                        if src_doc and src_doc not in backfill_source_docs and path_src != best_path:
                            backfill_source_docs.append(src_doc)
            if parsed_assets:
                parsed["assets"] = parsed_assets
            if parsed_years:
                parsed["years"] = sorted(parsed_years)
        if parsed:
            parsed["source_doc"] = str(best_path)
            if is_gpre_profile and backfill_source_docs:
                parsed["source_doc"] = (
                    f"{parsed['source_doc']}; annual Total assets backfilled from "
                    + "; ".join(backfill_source_docs)
                )
            parsed["source_qd"] = best_qd
        return parsed

    def _segment_revenue_valid_for_margin_bs(value_in: Any, *, segment_name: Any = "") -> bool:
        value_num = pd.to_numeric(value_in, errors="coerce")
        if pd.isna(value_num):
            return False
        value_float = float(value_num)
        if value_float <= 0:
            return False
        abs_value = abs(value_float)
        if abs_value < 1e-9:
            return False
        if is_pbi_profile:
            if abs_value >= 100_000.0:
                return abs_value >= 10_000_000.0
            return 10.0 <= abs_value < 1_000.0
        threshold = 1_000_000.0 if abs_value >= 100_000.0 else 0.1
        return abs_value >= threshold

    def _load_latest_quarterly_segment_data() -> Dict[str, Any]:
        if not enable_quarterly_segment_block:
            return {}
        def _parsed_quarterly_segments_from_slides_for_bs() -> Dict[str, Any]:
            if slides_segments is None or slides_segments.empty or "quarter" not in slides_segments.columns:
                return {}
            ss = slides_segments.copy()
            ss["quarter"] = pd.to_datetime(ss["quarter"], errors="coerce")
            if "value" not in ss.columns:
                return {}
            ss["value"] = pd.to_numeric(ss["value"], errors="coerce")
            ss = ss[ss["quarter"].notna() & ss["value"].notna()].copy()
            if is_anf_profile:
                ss = _filter_anf_quarterly_segment_actual_rows(
                    ss,
                    history_revenue_by_quarter=hist if isinstance(hist, pd.DataFrame) else None,
                )
            elif "period_type" in ss.columns:
                period_ser = ss["period_type"].astype(str).str.strip().str.lower()
                ss = ss[~period_ser.isin({"annual", "year", "fy", "ytd"})].copy()
            if ss.empty:
                return {}

            def _seg_label_local(seg_in: Any) -> str:
                seg_low = glx_normalize_text(str(seg_in or "")).lower()
                if "sending technology" in seg_low or re.search(r"\bsendtech\b", seg_low):
                    return "SendTech Solutions"
                if "presort" in seg_low:
                    return "Presort Services"
                if "total reportable" in seg_low:
                    return "Total reportable segments"
                return glx_normalize_text(str(seg_in or ""))

            metric_name_map = {
                "revenue": "Revenue",
                "adj_segment_ebit": "Adjusted EBIT",
                "adjusted segment ebit": "Adjusted EBIT",
                "adj_segment_da": "Depreciation & amortization",
                "adjusted segment da": "Depreciation & amortization",
                "adj_segment_ebitda": "Adjusted EBITDA",
                "adjusted segment ebitda": "Adjusted EBITDA",
            }
            dollar_metric_labels = {
                "Revenue",
                "Adjusted EBIT",
                "Depreciation & amortization",
                "Adjusted EBITDA",
            }

            def _normalize_slide_segment_value_for_bs(
                metric_label_in: str,
                value_in: Any,
                doc_txt_in: str,
            ) -> Optional[float]:
                value_num_in = pd.to_numeric(value_in, errors="coerce")
                if pd.isna(value_num_in):
                    return None
                value_float = float(value_num_in)
                if not is_pbi_profile or metric_label_in not in dollar_metric_labels:
                    return value_float
                abs_value = abs(value_float)
                if abs_value >= 100_000.0:
                    return value_float
                doc_low = str(doc_txt_in or "").lower()
                if "transcript" in doc_low or "metadata" in doc_low:
                    return None
                if 10.0 <= abs_value <= 10_000.0:
                    return value_float * 1_000_000.0
                return None

            rows_scored: List[Tuple[Tuple[float, float], Dict[str, Any]]] = []
            source_docs: List[str] = []
            for rec in ss.to_dict("records"):
                metric_key = str(rec.get("metric") or "").strip().lower()
                metric_label = metric_name_map.get(metric_key)
                seg_label = _seg_label_local(rec.get("segment"))
                if not metric_label or not seg_label:
                    continue
                value_num = pd.to_numeric(rec.get("value"), errors="coerce")
                if pd.isna(value_num):
                    continue
                doc_txt = str(rec.get("doc") or "").strip()
                normalized_value = _normalize_slide_segment_value_for_bs(metric_label, value_num, doc_txt)
                if normalized_value is None:
                    continue
                score = 0.0
                if "earnings_release" in str(rec.get("source") or "").lower():
                    score += 30.0
                if doc_txt.lower().endswith(".pdf"):
                    score += 45.0
                if pd.notna(rec.get("page")):
                    score += 15.0
                if "financial_statement" in doc_txt.lower():
                    score -= 20.0
                if abs(float(value_num)) >= 1_000_000.0:
                    score += 8.0
                elif abs(float(value_num)) > 0:
                    score -= 12.0
                rows_scored.append(
                    (
                        (score, abs(float(normalized_value))),
                        {
                            **rec,
                            "_metric_label": metric_label,
                            "_segment_label": seg_label,
                            "_normalized_value": float(normalized_value),
                        },
                    )
                )
                if doc_txt and doc_txt not in source_docs:
                    source_docs.append(doc_txt)

            store: Dict[str, Dict[str, Dict[pd.Timestamp, float]]] = {}
            for _score_tuple, rec in sorted(rows_scored, key=lambda item: item[0], reverse=True):
                metric_label = str(rec.get("_metric_label") or "")
                seg_label = str(rec.get("_segment_label") or "")
                q_ts = pd.Timestamp(rec.get("quarter"))
                bucket = store.setdefault(metric_label, {}).setdefault(seg_label, {})
                if q_ts not in bucket:
                    bucket[q_ts] = float(rec.get("_normalized_value"))

            if is_pbi_profile:
                store = _pbi_repair_total_reportable_segment_quarterly_totals_for_bs(store)

            revenue_map = dict(store.get("Revenue") or {})
            op_map = dict(store.get("Adjusted EBIT") or {})
            if revenue_map and op_map and "EBIT margin %" not in store:
                margin_map: Dict[str, Dict[pd.Timestamp, float]] = {}
                for seg_name, op_series in op_map.items():
                    rev_series = dict(revenue_map.get(seg_name) or {})
                    for q_key, op_val in dict(op_series or {}).items():
                        rev_val = pd.to_numeric(rev_series.get(q_key), errors="coerce")
                        op_num = pd.to_numeric(op_val, errors="coerce")
                        if pd.notna(op_num) and _segment_revenue_valid_for_margin_bs(rev_val, segment_name=seg_name):
                            margin_map.setdefault(seg_name, {})[q_key] = float(op_num) / float(rev_val)
                if margin_map:
                    store["EBIT margin %"] = margin_map
                    store["Segment operating margin %"] = margin_map

            if is_anf_profile:
                store = _anf_add_total_company_quarter_revenue_from_history(
                    store,
                    hist if isinstance(hist, pd.DataFrame) else None,
                )
                store = _anf_fill_brand_quarter_revenue_from_annual_segments_for_bs(
                    store,
                    slides_segments if isinstance(slides_segments, pd.DataFrame) else None,
                    hist if isinstance(hist, pd.DataFrame) else None,
                )
            quarters = sorted(
                {
                    pd.Timestamp(q).date()
                    for seg_map in store.values()
                    for q_map in seg_map.values()
                    for q in q_map.keys()
                }
            )
            if not store or not quarters:
                return {}
            return {
                "metrics": store,
                "quarters": quarters,
                "source_doc": " | ".join(source_docs[:3]) if source_docs else "Slides_Segments",
                "source_qd": max(quarters),
            }

        def _parsed_pbi_segment_release_tables_for_bs() -> Dict[str, Any]:
            if not is_pbi_profile:
                return {}

            def _parse_money_thousands(text_in: Any) -> Optional[float]:
                txt_num = str(text_in or "").strip()
                if not txt_num or txt_num in {"-", "\u2014", "\u2013"}:
                    return None
                neg = "(" in txt_num and ")" in txt_num
                txt_num = re.sub(r"[^0-9.\-]", "", txt_num)
                if not txt_num:
                    return None
                try:
                    val = float(txt_num)
                except Exception:
                    return None
                if neg:
                    val = -abs(val)
                return val * 1_000.0

            def _add_metric(
                store_in: Dict[str, Dict[str, Dict[pd.Timestamp, float]]],
                metric_name: str,
                segment_name: str,
                q_ts: pd.Timestamp,
                value_in: Any,
            ) -> None:
                value_num = _parse_money_thousands(value_in)
                if value_num is None:
                    return
                store_in.setdefault(metric_name, {}).setdefault(segment_name, {})[q_ts] = float(value_num)

            store: Dict[str, Dict[str, Dict[pd.Timestamp, float]]] = {}
            source_docs: List[str] = []
            for root in material_roots:
                er_dir = root / "earnings_release"
                if not er_dir.exists() or not er_dir.is_dir():
                    continue
                try:
                    files = sorted(er_dir.iterdir(), key=lambda pp: pp.name.lower())
                except Exception:
                    continue
                for path_in in files:
                    if not path_in.is_file() or path_in.suffix.lower() not in {".htm", ".html", ".txt"}:
                        continue
                    qd_local = (
                        source_infer_q_from_name(path_in.name)
                        or _parse_quarter_from_follow_text(path_in.name)
                        or _parse_quarter_from_filename(path_in.name)
                    )
                    if not isinstance(qd_local, date):
                        continue
                    try:
                        raw_txt = _read_operating_driver_text(path_in)
                    except Exception:
                        raw_txt = ""
                    if not raw_txt:
                        continue
                    txt = html.unescape(strip_html(raw_txt))
                    txt = re.sub(r"\s+", " ", txt).strip()
                    if "Business Segment Revenue" not in txt and "Adjusted Segment EBIT & EBITDA" not in txt:
                        continue
                    q_ts = pd.Timestamp(qd_local).normalize()
                    rev_pat = (
                        r"Business\s+Segment\s+Revenue.*?"
                        r"Sending\s+Technology\s+Solutions\s+\$?\s*([\(\)0-9,.\-]+)\s+\$?\s*[\(\)0-9,.\-]+.*?"
                        r"Presort\s+Services\s+\$?\s*([\(\)0-9,.\-]+)\s+\$?\s*[\(\)0-9,.\-]+.*?"
                        r"Total\s+revenue\s+\$?\s*([\(\)0-9,.\-]+)\s+\$?\s*[\(\)0-9,.\-]+"
                    )
                    rev_match = re.search(rev_pat, txt, flags=re.I | re.S)
                    if rev_match:
                        _add_metric(store, "Revenue", "SendTech Solutions", q_ts, rev_match.group(1))
                        _add_metric(store, "Revenue", "Presort Services", q_ts, rev_match.group(2))
                        _add_metric(store, "Revenue", "Total reportable segments", q_ts, rev_match.group(3))

                    ebit_pat = (
                        r"Adjusted\s+Segment\s+EBIT\s*&\s*EBITDA.*?"
                        r"Sending\s+Technology\s+Solutions\s+\$?\s*([\(\)0-9,.\-]+)\s+\$?\s*([\(\)0-9,.\-]+)\s+\$?\s*([\(\)0-9,.\-]+).*?"
                        r"Presort\s+Services\s+\$?\s*([\(\)0-9,.\-]+)\s+\$?\s*([\(\)0-9,.\-]+)\s+\$?\s*([\(\)0-9,.\-]+).*?"
                        r"Total\s+reportable\s+segments\s+\$?\s*([\(\)0-9,.\-]+)\s+\$?\s*([\(\)0-9,.\-]+)\s+\$?\s*([\(\)0-9,.\-]+)"
                    )
                    ebit_match = re.search(ebit_pat, txt, flags=re.I | re.S)
                    if ebit_match:
                        groups = list(ebit_match.groups())
                        for seg_name, offset in (
                            ("SendTech Solutions", 0),
                            ("Presort Services", 3),
                            ("Total reportable segments", 6),
                        ):
                            _add_metric(store, "Adjusted EBIT", seg_name, q_ts, groups[offset])
                            _add_metric(store, "Depreciation & amortization", seg_name, q_ts, groups[offset + 1])
                            _add_metric(store, "Adjusted EBITDA", seg_name, q_ts, groups[offset + 2])
                        _pbi_add_corporate_reconciliation_from_release_text(
                            store,
                            txt,
                            q_ts,
                            _parse_money_thousands,
                        )
                    if rev_match or ebit_match:
                        source_docs.append(str(path_in))

            if not store:
                return {}
            store = _pbi_repair_total_reportable_segment_quarterly_totals_for_bs(store)
            revenue_map = dict(store.get("Revenue") or {})
            op_map = dict(store.get("Adjusted EBIT") or {})
            if revenue_map and op_map:
                margin_map: Dict[str, Dict[pd.Timestamp, float]] = {}
                for seg_name, op_series in op_map.items():
                    rev_series = dict(revenue_map.get(seg_name) or {})
                    for q_key, op_val in dict(op_series or {}).items():
                        rev_val = pd.to_numeric(rev_series.get(q_key), errors="coerce")
                        op_num = pd.to_numeric(op_val, errors="coerce")
                        if pd.notna(op_num) and _segment_revenue_valid_for_margin_bs(rev_val, segment_name=seg_name):
                            margin_map.setdefault(seg_name, {})[q_key] = float(op_num) / float(rev_val)
                if margin_map:
                    store["EBIT margin %"] = margin_map
                    store["Segment operating margin %"] = margin_map
            quarters = sorted(
                {
                    pd.Timestamp(q).date()
                    for seg_map in store.values()
                    for q_map in seg_map.values()
                    for q in q_map.keys()
                }
            )
            return {
                "metrics": store,
                "quarters": quarters,
                "source_doc": " | ".join(dict.fromkeys(source_docs[-3:])),
                "source_qd": max(quarters) if quarters else None,
            }

        workbook_path = _latest_segment_financials_workbook()
        parsed: Dict[str, Any] = {}
        if workbook_path is not None:
            parsed = _parse_quarterly_segment_data_from_workbook(workbook_path)
            if parsed:
                if is_pbi_profile and parsed.get("metrics"):
                    parsed["metrics"] = _pbi_repair_total_reportable_segment_quarterly_totals_for_bs(parsed.get("metrics") or {})
                parsed["source_doc"] = str(workbook_path)
                parsed["source_qd"] = _parse_quarter_from_filename(workbook_path.name)

        def _merge_quarterly_segment_data(
            base: Dict[str, Any],
            overlay: Dict[str, Any],
        ) -> Dict[str, Any]:
            base_metrics = dict(base.get("metrics") or {})
            overlay_metrics = dict(overlay.get("metrics") or {})
            if not base_metrics:
                return overlay
            if not overlay_metrics:
                return base

            merged_metrics: Dict[str, Dict[str, Dict[pd.Timestamp, float]]] = {}

            def _copy_metrics(src_metrics: Dict[str, Any], *, replace_existing: bool) -> None:
                for metric_name, seg_map in dict(src_metrics or {}).items():
                    metric_bucket = merged_metrics.setdefault(str(metric_name), {})
                    for seg_name, q_map in dict(seg_map or {}).items():
                        seg_bucket = metric_bucket.setdefault(str(seg_name), {})
                        for q_key, value_in in dict(q_map or {}).items():
                            q_ts = pd.Timestamp(q_key)
                            value_num = pd.to_numeric(value_in, errors="coerce")
                            if pd.isna(value_num):
                                continue
                            if replace_existing or q_ts not in seg_bucket:
                                seg_bucket[q_ts] = float(value_num)

            _copy_metrics(base_metrics, replace_existing=True)
            _copy_metrics(overlay_metrics, replace_existing=False)
            if is_pbi_profile:
                merged_metrics = _pbi_repair_total_reportable_segment_quarterly_totals_for_bs(merged_metrics)
            if is_anf_profile:
                merged_metrics = _anf_add_total_company_quarter_revenue_from_history(
                    merged_metrics,
                    hist if isinstance(hist, pd.DataFrame) else None,
                )
                merged_metrics = _anf_fill_brand_quarter_revenue_from_annual_segments_for_bs(
                    merged_metrics,
                    slides_segments if isinstance(slides_segments, pd.DataFrame) else None,
                    hist if isinstance(hist, pd.DataFrame) else None,
                )

            quarters = sorted(
                {
                    pd.Timestamp(q).date()
                    for seg_map in merged_metrics.values()
                    for q_map in seg_map.values()
                    for q in q_map.keys()
                }
            )
            source_docs = [
                str(src).strip()
                for src in [base.get("source_doc"), overlay.get("source_doc")]
                if str(src or "").strip()
            ]
            return {
                "metrics": merged_metrics,
                "quarters": quarters,
                "source_doc": " | ".join(dict.fromkeys(source_docs)),
                "source_qd": max(quarters) if quarters else base.get("source_qd") or overlay.get("source_qd"),
            }

        metric_store = dict(parsed.get("metrics") or {})
        quarter_list = [
            pd.Timestamp(qd).date()
            for qd in list(parsed.get("quarters") or [])
            if isinstance(qd, (pd.Timestamp, date))
        ]
        latest_q = max(quarter_list) if quarter_list else None
        visible_latest_q = pd.Timestamp(max(qs)).date() if qs else None
        if metric_store and quarter_list and (visible_latest_q is None or latest_q == visible_latest_q):
            return parsed
        parsed_slides = _parsed_quarterly_segments_from_slides_for_bs()
        slide_qs = [
            pd.Timestamp(qd).date()
            for qd in list(parsed_slides.get("quarters") or [])
            if isinstance(qd, (pd.Timestamp, date))
        ]
        parsed_release = _parsed_pbi_segment_release_tables_for_bs()
        release_qs = [
            pd.Timestamp(qd).date()
            for qd in list(parsed_release.get("quarters") or [])
            if isinstance(qd, (pd.Timestamp, date))
        ]
        if parsed_release and release_qs and (visible_latest_q is None or max(release_qs) == visible_latest_q):
            merged = _merge_quarterly_segment_data(parsed, parsed_release) if metric_store and quarter_list else parsed_release
            if parsed_slides and slide_qs and (visible_latest_q is None or max(slide_qs) == visible_latest_q):
                merged = _merge_quarterly_segment_data(merged, parsed_slides)
            return merged
        if parsed_slides and slide_qs and metric_store and quarter_list and (visible_latest_q is None or max(slide_qs) == visible_latest_q):
            return _merge_quarterly_segment_data(parsed, parsed_slides)
        if parsed_slides and slide_qs and (visible_latest_q is None or max(slide_qs) == visible_latest_q):
            return parsed_slides
        return parsed if metric_store and quarter_list else {}

    def _pick_col(cands: List[str]) -> Optional[str]:
        cols_lc = {str(c).strip().lower(): c for c in h_idx.columns}
        for cand in cands:
            c = cols_lc.get(str(cand).strip().lower())
            if c is None:
                continue
            ser = pd.to_numeric(h_idx[c], errors="coerce")
            if ser.notna().any():
                return str(c)
        return None

    def _map_from_col(col: Optional[str]) -> Dict[pd.Timestamp, Optional[float]]:
        if col is None or col not in h_idx.columns:
            return {}
        ser = pd.to_numeric(h_idx[col], errors="coerce")
        out: Dict[pd.Timestamp, Optional[float]] = {}
        for k, v in ser.items():
            out[pd.Timestamp(k)] = float(v) if pd.notna(v) else None
        return out

    maps: Dict[str, Dict[pd.Timestamp, Optional[float]]] = {
        "cash": _map_from_col(_pick_col(["cash", "cash_and_cash_equivalents"])),
        "sti": _map_from_col(_pick_col(["short_term_investments", "short_term_investment"])),
        "ar": _map_from_col(_pick_col(["accounts_receivable", "accounts_receivable_net"])),
        "inventory": _map_from_col(_pick_col(["inventory"])),
        "assets_current": _map_from_col(_pick_col(["assets_current", "total_current_assets"])),
        "assets": _map_from_col(_pick_col(["assets", "total_assets"])),
        "ppne": _map_from_col(_pick_col(["ppne_net", "property_plant_equipment_net", "ppe_net"])),
        "goodwill": _map_from_col(_pick_col(["goodwill"])),
        "intangibles": _map_from_col(_pick_col(["intangibles", "intangible_assets_net_excluding_goodwill"])),
        "other_lt_assets": _map_from_col(_pick_col(["other_assets_noncurrent", "other_long_term_assets"])),
        "ap": _map_from_col(_pick_col(["accounts_payable", "accounts_payable_current"])),
        "accrued": _map_from_col(_pick_col(["accrued_liabilities", "accrued_liabilities_current", "other_accrued_liabilities"])),
        "def_rev": _map_from_col(_pick_col(["deferred_revenue", "deferred_revenue_current", "bank_deposits"])),
        "liabilities_current": _map_from_col(_pick_col(["liabilities_current", "total_current_liabilities"])),
        "liabilities": _map_from_col(_pick_col(["liabilities", "total_liabilities"])),
        "debt_st": _map_from_col(_pick_col(["debt_current", "short_term_debt", "long_term_debt_current"])),
        "debt_total": _map_from_col(_pick_col(["debt_core", "total_debt", "long_term_debt"])),
        "lease_lt": _map_from_col(_pick_col(["lease_liabilities_noncurrent", "lease_liabilities"])),
        "pension": _map_from_col(_pick_col(["pension_obligation_net", "pension_liabilities"])),
        "other_lt_liab": _map_from_col(_pick_col(["other_liabilities_noncurrent", "other_long_term_liabilities"])),
        "equity": _map_from_col(_pick_col(["total_equity", "stockholders_equity"])),
        "shares_out": _map_from_col(_pick_col(["shares_outstanding"])),
        "shares_dil": _map_from_col(_pick_col(["shares_diluted"])),
        "fin_recv": _map_from_col(_pick_col(["bank_finance_receivables"])),
        "deposits": _map_from_col(_pick_col(["bank_deposits"])),
        "bank_net_funding": _map_from_col(_pick_col(["bank_net_funding"])),
        "revenue": _map_from_col(_pick_col(["revenue"])),
    }

    local_bs_payloads = _shared_load_local_balance_sheet_detail_payloads({pd.Timestamp(x).date() for x in qs})

    def _payload_map(keys: List[str], hist_cols: Optional[List[str]] = None) -> Dict[pd.Timestamp, Optional[float]]:
        hist_map = _map_from_col(_pick_col(hist_cols or [])) if hist_cols else {}
        out_map: Dict[pd.Timestamp, Optional[float]] = {}
        for q in [pd.Timestamp(x) for x in qs]:
            hist_val = hist_map.get(q)
            if hist_val is not None:
                out_map[q] = hist_val
                continue
            payload = local_bs_payloads.get(q.date()) or {}
            vals = payload.get("values", {}) or {}
            chosen: Optional[float] = None
            for key_name in keys:
                if key_name in vals and vals.get(key_name) is not None:
                    chosen = float(vals[key_name])
                    break
            out_map[q] = chosen
        return out_map

    maps["restricted_cash"] = _payload_map(["restricted_cash"], ["restricted_cash"])
    maps["prepaid_expenses_other_current"] = _payload_map(
        ["prepaid_other_current_assets", "prepaid_expenses_and_other_current_assets"],
        ["prepaid_other_current_assets"],
    )
    maps["derivative_asset"] = _payload_map(["derivative_assets"], ["derivative_assets"])
    maps["rou_asset"] = _payload_map(["operating_lease_rou_assets"], ["operating_lease_rou_assets"])
    maps["deferred_tax"] = _payload_map(["deferred_income_taxes_net"], ["deferred_income_taxes_net"])
    maps["goodwill"] = _payload_map(["goodwill"], ["goodwill"])
    maps["intangibles"] = _payload_map(
        ["intangibles", "intangible_assets_net_excluding_goodwill"],
        ["intangibles", "intangible_assets_net_excluding_goodwill"],
    )
    maps["short_term_notes_borrowings"] = _payload_map(
        ["short_term_notes_payable_and_other_borrowings"],
        ["short_term_notes_payable_and_other_borrowings"],
    )
    maps["current_maturities_ltd"] = _payload_map(
        ["current_maturities_of_long_term_debt"],
        ["current_maturities_of_long_term_debt", "long_term_debt_current", "debt_current"],
    )
    maps["derivative_liability"] = _payload_map(["derivative_liabilities"], ["derivative_liabilities"])
    maps["lease_current"] = _payload_map(["operating_lease_current_liabilities"], ["operating_lease_current_liabilities"])
    maps["lease_long_term"] = _payload_map(
        ["operating_lease_long_term_liabilities"],
        ["operating_lease_long_term_liabilities", "lease_liabilities_noncurrent"],
    )
    maps["carbon_equipment_liabilities"] = _payload_map(["carbon_equipment_liabilities"], ["carbon_equipment_liabilities"])
    maps["other_liabilities_explicit"] = _payload_map(["other_liabilities"], ["other_liabilities_noncurrent"])
    for q in [pd.Timestamp(x) for x in qs]:
        gw_v = maps["goodwill"].get(q)
        if gw_v is not None and abs(float(gw_v)) < 1_000_000.0:
            maps["goodwill"][q] = None
        int_v = maps["intangibles"].get(q)
        if int_v is not None and abs(float(int_v)) < 1_000_000.0:
            maps["intangibles"][q] = None
    maps["goodwill"] = _carry_forward_low_change_series(maps["goodwill"], list(qs))
    maps["intangibles"] = _carry_forward_low_change_series(maps["intangibles"], list(qs))

    if is_anf_profile:
        marketable_source_map = _map_from_col(
            _pick_col(["marketable_securities", "short_term_investments", "short_term_investment"])
        )
        lease_total_source_map = _map_from_col(
            _pick_col(["lease_liabilities", "operating_lease_liabilities", "lease_liabilities_total"])
        )
        marketable_map_norm = {pd.Timestamp(k).normalize(): v for k, v in dict(marketable_source_map or {}).items()}
        lease_total_map_norm = {pd.Timestamp(k).normalize(): v for k, v in dict(lease_total_source_map or {}).items()}
        for q in [pd.Timestamp(x).normalize() for x in qs]:
            if maps["sti"].get(q) is None and marketable_map_norm.get(q) is not None:
                maps["sti"][q] = float(marketable_map_norm[q])

            lease_total_v = lease_total_map_norm.get(q)
            lease_current_v = maps["lease_current"].get(q)
            lease_long_v = maps["lease_long_term"].get(q)
            if lease_long_v is None:
                lease_long_v = maps["lease_lt"].get(q)
            if lease_total_v is not None:
                if lease_current_v is None and lease_long_v is not None:
                    maps["lease_current"][q] = max(float(lease_total_v) - float(lease_long_v), 0.0)
                if maps["lease_long_term"].get(q) is None and lease_current_v is not None:
                    maps["lease_long_term"][q] = max(float(lease_total_v) - float(lease_current_v), 0.0)

            if maps["liabilities"].get(q) is None:
                assets_v = maps["assets"].get(q)
                equity_v = maps["equity"].get(q)
                if assets_v is not None and equity_v is not None:
                    derived_liabilities = float(assets_v) - float(equity_v)
                    if derived_liabilities >= -1_000_000.0:
                        maps["liabilities"][q] = derived_liabilities

    total_cash_restricted_map: Dict[pd.Timestamp, Optional[float]] = {}
    for q in [pd.Timestamp(x) for x in qs]:
        cash_v = maps["cash"].get(q)
        restricted_v = maps["restricted_cash"].get(q)
        if cash_v is None and restricted_v is None:
            total_cash_restricted_map[q] = None
        else:
            total_cash_restricted_map[q] = float(cash_v or 0.0) + float(restricted_v or 0.0)

    def _qoq_delta_map(src_map: Dict[pd.Timestamp, Optional[float]]) -> Dict[pd.Timestamp, Optional[float]]:
        out_map: Dict[pd.Timestamp, Optional[float]] = {}
        prev_val: Optional[float] = None
        for q in [pd.Timestamp(x) for x in qs]:
            vv = src_map.get(q)
            if vv is None or prev_val is None:
                out_map[q] = None
            else:
                out_map[q] = float(vv) - float(prev_val)
            prev_val = float(vv) if vv is not None else None
        return out_map

    delta_cash_qoq_map = _qoq_delta_map(maps["cash"])
    delta_total_debt_qoq_map = _qoq_delta_map(maps["debt_total"])

    accrued_tag_by_q: Dict[pd.Timestamp, str] = {}
    if audit is not None and not audit.empty:
        try:
            aud_bs = audit.copy()
            qcol_a = _resolve_col(aud_bs, ["quarter", "period_end"])
            mcol_a = _resolve_col(aud_bs, ["metric"])
            tcol_a = _resolve_col(aud_bs, ["tag"])
            vcol_a = _resolve_col(aud_bs, ["value"])
            scol_a = _resolve_col(aud_bs, ["source"])
            fcol_a = _resolve_col(aud_bs, ["filed", "filed_date"])
            if qcol_a and mcol_a and tcol_a and vcol_a:
                aud_bs["_q"] = pd.to_datetime(aud_bs[qcol_a], errors="coerce").dt.normalize()
                if fcol_a:
                    aud_bs["_filed"] = pd.to_datetime(aud_bs[fcol_a], errors="coerce")
                else:
                    aud_bs["_filed"] = pd.NaT
                aud_bs["_metric"] = aud_bs[mcol_a].astype(str).str.strip().str.lower()
                if scol_a:
                    aud_bs["_source"] = aud_bs[scol_a].astype(str).str.strip().str.lower()
                    aud_bs = aud_bs[aud_bs["_source"] != "missing"]
                aud_bs = aud_bs[aud_bs["_q"].notna()]
                if not aud_bs.empty:
                    aud_bs["_val_num"] = pd.to_numeric(aud_bs[vcol_a], errors="coerce")
                    aud_bs = aud_bs[aud_bs["_val_num"].notna()]
                    accrued_metrics = {"accrued_liabilities_current", "accrued_liabilities", "other_accrued_liabilities"}
                    a_acc = aud_bs[aud_bs["_metric"].isin(accrued_metrics)].copy()
                    if not a_acc.empty:
                        a_acc = a_acc.sort_values("_filed")
                        for _, rr in a_acc.iterrows():
                            qk = pd.Timestamp(rr["_q"])
                            accrued_tag_by_q[qk] = str(rr[tcol_a] or "").strip().lower()
        except Exception:
            pass

    prepaid_map: Dict[pd.Timestamp, Optional[float]] = {}
    for q in [pd.Timestamp(x) for x in qs]:
        explicit_prepaid = maps["prepaid_expenses_other_current"].get(q)
        if explicit_prepaid is not None:
            prepaid_map[q] = float(explicit_prepaid)
            continue
        ac = maps["assets_current"].get(q)
        if ac is None:
            prepaid_map[q] = None
            continue
        parts = [maps["cash"].get(q), maps["sti"].get(q), maps["ar"].get(q), maps["inventory"].get(q), maps["derivative_asset"].get(q)]
        if any(v is None for v in parts):
            prepaid_map[q] = None
            continue
        prepaid_map[q] = float(ac) - sum(float(v) for v in parts if v is not None)

    debt_lt_map: Dict[pd.Timestamp, Optional[float]] = {}
    for q in [pd.Timestamp(x) for x in qs]:
        total = maps["debt_total"].get(q)
        st = maps["current_maturities_ltd"].get(q)
        if st is None:
            st = maps["debt_st"].get(q)
        if total is None:
            debt_lt_map[q] = None
        elif st is None:
            debt_lt_map[q] = float(total)
        else:
            debt_lt_map[q] = float(total) - float(st)

    # If accrued liabilities came from AP+accrued combined tag, strip AP where both are available.
    for q in [pd.Timestamp(x) for x in qs]:
        tag_l = str(accrued_tag_by_q.get(q) or "").strip().lower()
        if "accountspayableandaccruedliabilitiescurrent" not in tag_l:
            continue
        acc_v = maps["accrued"].get(q)
        ap_v = maps["ap"].get(q)
        if acc_v is None or ap_v is None:
            continue
        adj_v = float(acc_v) - float(ap_v)
        if adj_v >= -1_000_000.0:
            maps["accrued"][q] = max(adj_v, 0.0)

    # Residual fallbacks for long-term "other" buckets when explicit tags are absent.
    for q in [pd.Timestamp(x) for x in qs]:
        if maps["other_lt_assets"].get(q) is None:
            a_tot = maps["assets"].get(q)
            a_cur = maps["assets_current"].get(q)
            ppne_v = maps["ppne"].get(q)
            rou_v = maps["rou_asset"].get(q)
            dtx_v = maps["deferred_tax"].get(q)
            gw_v = maps["goodwill"].get(q)
            int_v = maps["intangibles"].get(q)
            if all(v is not None for v in [a_tot, a_cur, ppne_v, gw_v, int_v]):
                resid_a = float(a_tot) - (
                    float(a_cur)
                    + float(ppne_v)
                    + float(rou_v or 0.0)
                    + float(dtx_v or 0.0)
                    + float(gw_v)
                    + float(int_v)
                )
                if resid_a >= -5_000_000.0 and abs(resid_a) <= max(abs(float(a_tot)) * 1.25, 10_000_000.0):
                    maps["other_lt_assets"][q] = resid_a
                elif q == pd.Timestamp(qs[-1]):
                    qa_rows.append(
                        {
                            "quarter": pd.Timestamp(q).date(),
                            "metric": "BS_Segments",
                            "check": "other_lt_assets_residual",
                            "status": "warn",
                            "message": f"Residual other LT assets looks unstable ({resid_a/1e6:,.1f}m); left blank.",
                            "source": "History_Q",
                        }
                    )
        if maps["ppne"].get(q) is None:
            a_tot = maps["assets"].get(q)
            a_cur = maps["assets_current"].get(q)
            oth_a = maps["other_lt_assets"].get(q)
            rou_v = maps["rou_asset"].get(q)
            dtx_v = maps["deferred_tax"].get(q)
            gw_v = maps["goodwill"].get(q)
            int_v = maps["intangibles"].get(q)
            if all(v is not None for v in [a_tot, a_cur, oth_a, gw_v, int_v]):
                resid_ppne = float(a_tot) - (
                    float(a_cur)
                    + float(oth_a)
                    + float(rou_v or 0.0)
                    + float(dtx_v or 0.0)
                    + float(gw_v)
                    + float(int_v)
                )
                if resid_ppne >= -5_000_000.0 and abs(resid_ppne) <= max(abs(float(a_tot)) * 1.25, 10_000_000.0):
                    maps["ppne"][q] = resid_ppne
                elif q == pd.Timestamp(qs[-1]):
                    qa_rows.append(
                        {
                            "quarter": pd.Timestamp(q).date(),
                            "metric": "BS_Segments",
                            "check": "ppne_residual",
                            "status": "warn",
                            "message": f"Residual PP&E looks unstable ({resid_ppne/1e6:,.1f}m); left blank.",
                            "source": "History_Q",
                        }
                    )
        if maps["other_lt_liab"].get(q) is None:
            l_tot = maps["liabilities"].get(q)
            l_cur = maps["liabilities_current"].get(q)
            d_lt = debt_lt_map.get(q)
            lease_v = maps["lease_long_term"].get(q) if maps["lease_long_term"].get(q) is not None else maps["lease_lt"].get(q)
            pension_v = maps["pension"].get(q)
            carbon_v = maps["carbon_equipment_liabilities"].get(q)
            explicit_other_v = maps["other_liabilities_explicit"].get(q)
            if explicit_other_v is not None:
                maps["other_lt_liab"][q] = float(explicit_other_v)
            elif all(v is not None for v in [l_tot, l_cur, d_lt, lease_v, pension_v]):
                resid_l = float(l_tot) - (
                    float(l_cur)
                    + float(d_lt)
                    + float(lease_v)
                    + float(pension_v)
                    + float(carbon_v or 0.0)
                )
                if resid_l >= -5_000_000.0 and abs(resid_l) <= max(abs(float(l_tot)) * 1.25, 10_000_000.0):
                    maps["other_lt_liab"][q] = resid_l
                elif q == pd.Timestamp(qs[-1]):
                    qa_rows.append(
                        {
                            "quarter": pd.Timestamp(q).date(),
                            "metric": "BS_Segments",
                            "check": "other_lt_liabilities_residual",
                            "status": "warn",
                            "message": f"Residual other LT liabilities looks unstable ({resid_l/1e6:,.1f}m); left blank.",
                            "source": "History_Q",
                        }
                    )

    numeric_rows: List[int] = []
    pct_rows: List[int] = []
    no_bucket_rows: set = set()
    delta_nwc_rows: set = set()
    nwc_rows: set = set()
    row_color_contexts: Dict[int, Tuple[str, str]] = {}
    row_hidden_source_values: Dict[int, Dict[pd.Timestamp, float]] = {}
    row_visible_keys: Dict[int, List[Any]] = {}

    def _apply_data_font(cell: Any, *, bold_flag: bool = False) -> None:
        f0 = cell.font if cell.font is not None else Font()
        cell.font = Font(
            name=f0.name,
            size=valuation_data_font_size,
            bold=bool(bold_flag),
            italic=f0.italic,
            color=f0.color,
            underline=f0.underline,
            strike=f0.strike,
        )

    def _write_section(label: str) -> None:
        nonlocal row_idx
        ws[f"A{row_idx}"] = label
        if valuation_section_label_style is not None:
            for cc in range(1, last_col + 1):
                c = ws[f"{get_column_letter(cc)}{row_idx}"]
                if cc == 1:
                    c._style = copy(valuation_section_label_style)
                elif valuation_section_col_style is not None:
                    c._style = copy(valuation_section_col_style)
                else:
                    c._style = copy(valuation_section_label_style)
                if cc != 1:
                    c.value = None
        else:
            ws[f"A{row_idx}"].font = bold
            for cc in range(1, last_col + 1):
                c = ws[f"{get_column_letter(cc)}{row_idx}"]
                c.fill = section_fill
                c.border = thin_border
        row_idx += 1

    def _promote_section_band(section_row: int) -> None:
        for cc in range(1, last_col + 1):
            c = ws.cell(row=section_row, column=cc)
            c.fill = copy(title_fill)
            c.border = thin_border
            if cc == 1 or c.value not in (None, ""):
                c.font = Font(
                    bold=True,
                    size=float(getattr(c.font, "size", header_size) or header_size),
                    color="FFFFFF",
                )

    def _set_comment(cell: Any, txt: str) -> None:
        if not txt:
            return
        try:
            _set_cell_comment_local(cell, txt)
        except Exception:
            pass

    def _write_metric(
        label: str,
        val_map: Dict[pd.Timestamp, Optional[float]],
        *,
        value_scale: float = 1e6,
        number_format: str = "#,##0.0",
        bold_label: bool = False,
        comments: Optional[Dict[pd.Timestamp, str]] = None,
        bucket_mode: str = "qoq",
    ) -> int:
        nonlocal row_idx
        ws[f"A{row_idx}"] = label
        if valuation_label_style is not None:
            ws[f"A{row_idx}"]._style = copy(valuation_label_style)
            _apply_data_font(ws[f"A{row_idx}"], bold_flag=False)
        else:
            ws[f"A{row_idx}"].font = Font(bold=False, size=valuation_data_font_size)
        metric_row = row_idx
        for i, q in enumerate(qs):
            qk = pd.Timestamp(q)
            val = val_map.get(qk)
            cc = start_col + i
            if val is None:
                cell_value = None
            else:
                cell_value = float(val) / value_scale if value_scale != 1.0 else float(val)
                if is_anf_profile:
                    fmt_txt = str(number_format or "")
                    if "%" in fmt_txt:
                        cell_value = round(float(cell_value), 4)
                    elif "0.00x" in fmt_txt:
                        cell_value = round(float(cell_value), 2)
                    elif fmt_txt == "0.0":
                        cell_value = round(float(cell_value), 1)
                    else:
                        cell_value = round(float(cell_value), 1)
            cell = ws.cell(row=row_idx, column=cc, value=cell_value)
            if valuation_numeric_style is not None:
                cell._style = copy(valuation_numeric_style)
            _apply_data_font(cell, bold_flag=False)
            cell.number_format = number_format
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")
            if comments and qk in comments:
                _set_comment(cell, comments[qk])
        row_hidden_source_values[metric_row] = {
            pd.Timestamp(src_q).to_period("Q").end_time.normalize(): (
                float(src_val) / value_scale if value_scale != 1.0 else float(src_val)
            )
            for src_q, src_val in dict(val_map or {}).items()
            if src_val is not None and pd.notna(pd.to_numeric(src_val, errors="coerce"))
        }
        row_visible_keys[metric_row] = [pd.Timestamp(q).normalize() for q in qs]
        numeric_rows.append(metric_row)
        if "%" in str(number_format):
            pct_rows.append(metric_row)
        if bucket_mode == "none":
            no_bucket_rows.add(metric_row)
        elif bucket_mode == "delta_nwc":
            delta_nwc_rows.add(metric_row)
        elif bucket_mode == "nwc":
            nwc_rows.add(metric_row)
        row_idx += 1
        return metric_row

    _write_section("Liquidity / Assets")
    _write_metric("Cash & cash equivalents", maps["cash"])
    _write_metric("Restricted cash", maps["restricted_cash"])
    _write_metric("Total cash + restricted cash", total_cash_restricted_map, bold_label=True, bucket_mode="none")
    _write_metric("Δ Cash QoQ ($m)", delta_cash_qoq_map, bucket_mode="none")
    _write_metric("Short-term investments", maps["sti"])
    _write_metric("Accounts receivable (net)", maps["ar"])
    _write_metric("Inventory", maps["inventory"])
    _write_metric("Derivative financial instruments (asset)", maps["derivative_asset"])
    _write_metric("Prepaid & other current assets", prepaid_map)
    _write_metric("Total current assets", maps["assets_current"], bold_label=True)
    _write_metric("Property, plant & equipment (net)", maps["ppne"])
    _write_metric("Operating lease right-of-use assets", maps["rou_asset"])
    _write_metric("Deferred income taxes, net", maps["deferred_tax"])
    _write_metric("Goodwill", maps["goodwill"])
    _write_metric("Intangibles (net)", maps["intangibles"])
    _write_metric("Other long-term assets", maps["other_lt_assets"])
    _write_metric("Total assets", maps["assets"], bold_label=True)
    goodwill_pct_assets_map: Dict[pd.Timestamp, Optional[float]] = {}
    for q in [pd.Timestamp(x) for x in qs]:
        gw_v = maps["goodwill"].get(q)
        assets_v = maps["assets"].get(q)
        if gw_v is None or assets_v in (None, 0):
            goodwill_pct_assets_map[q] = None
        else:
            goodwill_pct_assets_map[q] = float(gw_v) / float(assets_v)
    _write_metric("Goodwill % of assets", goodwill_pct_assets_map, value_scale=1.0, number_format="0.0%", bucket_mode="none")

    _write_section("Liabilities")
    _write_metric("Accounts payable", maps["ap"])
    _write_metric("Accrued liabilities", maps["accrued"])
    _write_metric("Deferred revenue / customer deposits", maps["def_rev"])
    _write_metric("Derivative financial instruments (liability)", maps["derivative_liability"])
    _write_metric("Short-term notes payable and other borrowings", maps["short_term_notes_borrowings"])
    _write_metric("Current maturities of long-term debt", maps["current_maturities_ltd"])
    _write_metric("Operating lease current liabilities", maps["lease_current"])
    _write_metric("Total current liabilities", maps["liabilities_current"], bold_label=True)
    nwc_map: Dict[pd.Timestamp, Optional[float]] = {}
    delta_nwc_map: Dict[pd.Timestamp, Optional[float]] = {}
    current_ratio_map: Dict[pd.Timestamp, Optional[float]] = {}
    quick_ratio_map: Dict[pd.Timestamp, Optional[float]] = {}
    prev_nwc_val: Optional[float] = None
    for q in [pd.Timestamp(x) for x in qs]:
        ca_v = maps["assets_current"].get(q)
        cl_v = maps["liabilities_current"].get(q)
        c_v = maps["cash"].get(q)
        sti_v = maps["sti"].get(q)
        ar_v = maps["ar"].get(q)
        if ca_v is not None and cl_v is not None:
            nwc_v = float(ca_v) - float(cl_v)
            nwc_map[q] = nwc_v
            if prev_nwc_val is None:
                delta_nwc_map[q] = None
            else:
                delta_nwc_map[q] = nwc_v - float(prev_nwc_val)
            prev_nwc_val = nwc_v
            if abs(float(cl_v)) > 1e-12:
                current_ratio_map[q] = float(ca_v) / float(cl_v)
            else:
                current_ratio_map[q] = None
        else:
            nwc_map[q] = None
            delta_nwc_map[q] = None
            current_ratio_map[q] = None
            prev_nwc_val = None
        quick_num = None
        if c_v is not None and ar_v is not None:
            quick_num = float(c_v) + float(sti_v or 0.0) + float(ar_v)
        if quick_num is not None and cl_v is not None and abs(float(cl_v)) > 1e-12:
            quick_ratio_map[q] = quick_num / float(cl_v)
        else:
            quick_ratio_map[q] = None
    _write_metric("Net working capital", nwc_map, bold_label=True, bucket_mode="nwc")
    _write_metric("Δ NWC QoQ", delta_nwc_map, bucket_mode="delta_nwc")
    _write_metric("Current ratio", current_ratio_map, value_scale=1.0, number_format="0.00x", bucket_mode="none")
    _write_metric("Quick ratio", quick_ratio_map, value_scale=1.0, number_format="0.00x", bucket_mode="none")
    _write_metric("Long-term debt", debt_lt_map)
    _write_metric("Δ Total debt QoQ ($m)", delta_total_debt_qoq_map, bucket_mode="none")
    if _should_render_carbon_equipment_liabilities(ticker, maps["carbon_equipment_liabilities"]):
        _write_metric("Carbon equipment liabilities", maps["carbon_equipment_liabilities"])
    _write_metric(
        "Operating lease long-term liabilities",
        {q: (maps["lease_long_term"].get(q) if maps["lease_long_term"].get(q) is not None else maps["lease_lt"].get(q)) for q in [pd.Timestamp(x) for x in qs]},
    )
    _write_metric("Pension / OPEB obligation (net)", maps["pension"])
    _write_metric("Other long-term liabilities", maps["other_lt_liab"])
    _write_metric("Total liabilities", maps["liabilities"], bold_label=True)

    _write_section("Equity")
    _write_metric("Total equity", maps["equity"], bold_label=True)
    _write_metric("Shares outstanding (m)", maps["shares_out"], value_scale=1e6, number_format="0.0")
    _write_metric("Shares diluted (m)", maps["shares_dil"], value_scale=1e6, number_format="0.0")

    if is_anf_profile:
        qts = [pd.Timestamp(x) for x in qs]
        all_qts = [pd.Timestamp(x) for x in sorted(h_idx.index)]
        all_q_index = {pd.Timestamp(q): idx for idx, q in enumerate(all_qts)}

        def _yoy_map(src_map: Dict[pd.Timestamp, Optional[float]]) -> Dict[pd.Timestamp, Optional[float]]:
            out: Dict[pd.Timestamp, Optional[float]] = {}
            for q in qts:
                cur = src_map.get(q)
                idx_q = all_q_index.get(q)
                prior_q = all_qts[idx_q - 4] if idx_q is not None and idx_q >= 4 else None
                prior = src_map.get(prior_q) if prior_q is not None else None
                if cur is None or prior is None or abs(float(prior)) < 1e-9:
                    out[q] = None
                else:
                    out[q] = (float(cur) - float(prior)) / abs(float(prior))
            return out

        inventory_yoy_map = _yoy_map(maps["inventory"])
        sales_yoy_map = _yoy_map(maps["revenue"])
        inventory_vs_sales_map: Dict[pd.Timestamp, Optional[float]] = {}
        share_count_yoy_map = _yoy_map(maps["shares_dil"])
        total_lease_map: Dict[pd.Timestamp, Optional[float]] = {}
        net_cash_map: Dict[pd.Timestamp, Optional[float]] = {}
        for q in qts:
            inv_yoy = inventory_yoy_map.get(q)
            sales_yoy = sales_yoy_map.get(q)
            inventory_vs_sales_map[q] = (
                None if inv_yoy is None or sales_yoy is None else float(inv_yoy) - float(sales_yoy)
            )
            lease_current = maps["lease_current"].get(q) or 0.0
            lease_long = maps["lease_long_term"].get(q)
            if lease_long is None:
                lease_long = maps["lease_lt"].get(q)
            total_lease_map[q] = None if lease_long is None and not lease_current else float(lease_current or 0.0) + float(lease_long or 0.0)
            cash_v = maps["cash"].get(q)
            sti_v = maps["sti"].get(q) or 0.0
            debt_v = maps["debt_total"].get(q) or 0.0
            net_cash_map[q] = None if cash_v is None else float(cash_v) + float(sti_v) - float(debt_v)

        _write_section("ANF retail BS drivers")
        _write_metric("Inventory YoY", inventory_yoy_map, value_scale=1.0, number_format="0.0%", bucket_mode="none")
        _write_metric("Sales YoY", sales_yoy_map, value_scale=1.0, number_format="0.0%", bucket_mode="none")
        _write_metric("Inventory less sales YoY", inventory_vs_sales_map, value_scale=1.0, number_format="0.0%", bucket_mode="none")
        _write_metric("Net cash", net_cash_map, bucket_mode="none")
        _write_metric("Total lease liabilities", total_lease_map, bucket_mode="none")
        _write_metric("Diluted shares YoY", share_count_yoy_map, value_scale=1.0, number_format="0.0%", bucket_mode="none")

    if bank_metrics_enabled:
        _write_section("Bank / Financing (optional)")
        _write_metric("Finance receivables (total)", maps["fin_recv"])
        _write_metric("Deposits (bank/customer)", maps["deposits"])
        _write_metric("Bank net funding", maps["bank_net_funding"])

    quarterly_segment_data = _load_latest_quarterly_segment_data()
    quarterly_metrics = dict(quarterly_segment_data.get("metrics") or {})
    if quarterly_metrics and "Revenue" in quarterly_metrics:
        revenue_for_margin = dict(quarterly_metrics.get("Revenue") or {})
        for margin_metric in ("EBIT margin %", "Segment operating margin %"):
            margin_metric_map = dict(quarterly_metrics.get(margin_metric) or {})
            if not margin_metric_map:
                continue
            filtered_margin_map: Dict[str, Dict[pd.Timestamp, float]] = {}
            for seg_name, q_map in margin_metric_map.items():
                rev_series = dict(revenue_for_margin.get(seg_name) or {})
                for q_key, margin_val in dict(q_map or {}).items():
                    q_ts = pd.Timestamp(q_key)
                    if _segment_revenue_valid_for_margin_bs(rev_series.get(q_ts), segment_name=seg_name):
                        filtered_margin_map.setdefault(seg_name, {})[q_ts] = margin_val
            if filtered_margin_map:
                quarterly_metrics[margin_metric] = filtered_margin_map
            else:
                quarterly_metrics.pop(margin_metric, None)
    if "EBIT margin %" in quarterly_metrics:
        ebit_margin_metrics = dict(quarterly_metrics.get("EBIT margin %") or {})
        segment_margin_metrics = dict(quarterly_metrics.get("Segment operating margin %") or {})
        if segment_margin_metrics:
            for seg_name, q_map in ebit_margin_metrics.items():
                seg_bucket = segment_margin_metrics.setdefault(seg_name, {})
                for q_key, margin_val in dict(q_map or {}).items():
                    q_ts = pd.Timestamp(q_key)
                    if q_ts not in seg_bucket:
                        seg_bucket[q_ts] = margin_val
        else:
            segment_margin_metrics = ebit_margin_metrics
        if segment_margin_metrics:
            quarterly_metrics["Segment operating margin %"] = segment_margin_metrics
    quarterly_source_note = str(quarterly_segment_data.get("source_doc") or "").strip()
    if is_anf_profile:
        quarterly_metrics = _anf_add_total_company_quarter_revenue_from_history(
            quarterly_metrics,
            maps["revenue"],
            qs,
        )
    if quarterly_metrics:
        row_idx += 1
        quarterly_section_row = row_idx
        _write_section("Quarterly segments")
        _promote_section_band(quarterly_section_row)
        ws.row_dimensions[quarterly_section_row].height = 24.0
        if is_anf_profile:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=last_col)
            note_cell = ws.cell(row=row_idx, column=1, value=ANF_SEGMENT_BRAND_EXPLANATION)
            note_cell.font = Font(italic=True, size=10, color="666666")
            note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            for cc in range(1, last_col + 1):
                ws.cell(row=row_idx, column=cc).fill = copy(section_fill)
                ws.cell(row=row_idx, column=cc).border = thin_border
            ws.row_dimensions[row_idx].height = 24.0
            row_idx += 1

        quarterly_metric_order = [
            ("Revenue", "#,##0.0", 1e6),
            ("Adjusted EBIT", "#,##0.0", 1e6),
            ("Segment operating margin %", "0.0%", 1.0),
            ("EBIT margin %", "0.0%", 1.0),
            ("Depreciation & amortization", "#,##0.0", 1e6),
            ("Adjusted EBITDA", "#,##0.0", 1e6),
        ]

        def _write_quarterly_segment_metric(
            metric_label: str,
            seg_values: Dict[str, Dict[pd.Timestamp, float]],
            *,
            number_format: str,
            value_scale: float,
        ) -> None:
            nonlocal row_idx
            eligible_segments: List[str] = []
            ordered_segments = list(quarterly_segment_labels) or list(seg_values.keys())
            for seg in ordered_segments + [s for s in seg_values.keys() if s not in ordered_segments]:
                if seg in eligible_segments or seg not in seg_values:
                    continue
                has_value = any(
                    vv is not None and abs(float(vv)) > 1e-9
                    for q in [pd.Timestamp(x) for x in qs]
                    for vv in [seg_values.get(seg, {}).get(pd.Timestamp(q))]
                )
                if has_value:
                    eligible_segments.append(seg)
            if not eligible_segments:
                return
            _write_section(metric_label)
            for seg_idx, seg in enumerate(eligible_segments):
                ws[f"A{row_idx}"] = seg
                if valuation_label_style is not None:
                    ws[f"A{row_idx}"]._style = copy(valuation_label_style)
                    _apply_data_font(ws[f"A{row_idx}"], bold_flag=False)
                else:
                    ws[f"A{row_idx}"].font = Font(bold=False, size=valuation_data_font_size)
                for idx_q, q in enumerate(qs):
                    cc = start_col + idx_q
                    qk = pd.Timestamp(q)
                    vv = seg_values.get(seg, {}).get(qk)
                    if vv is None:
                        cell_value = None
                    else:
                        cell_value = float(vv) / value_scale if value_scale != 1.0 else float(vv)
                        if is_anf_profile:
                            if "%" in str(number_format):
                                cell_value = round(float(cell_value), 3)
                            else:
                                cell_value = round(float(cell_value), 1)
                    cell = ws.cell(row=row_idx, column=cc, value=cell_value)
                    if valuation_numeric_style is not None:
                        cell._style = copy(valuation_numeric_style)
                    _apply_data_font(cell, bold_flag=False)
                    cell.number_format = number_format
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="right")
                    if quarterly_source_note and idx_q == 0:
                        _set_comment(cell, quarterly_source_note)
                row_hidden_source_values[row_idx] = {
                    pd.Timestamp(src_q).to_period("Q").end_time.normalize(): (
                        float(src_val) / value_scale if value_scale != 1.0 else float(src_val)
                    )
                    for src_q, src_val in dict(seg_values.get(seg, {}) or {}).items()
                    if src_val is not None and pd.notna(pd.to_numeric(src_val, errors="coerce"))
                }
                row_visible_keys[row_idx] = [pd.Timestamp(q).normalize() for q in qs]
                numeric_rows.append(row_idx)
                row_color_contexts[row_idx] = ("Quarterly segments", metric_label)
                if "%" in str(number_format):
                    pct_rows.append(row_idx)
                row_idx += 1
                next_seg = eligible_segments[seg_idx + 1] if seg_idx + 1 < len(eligible_segments) else ""
                if is_anf_profile and seg == "APAC" and next_seg == "Total Company":
                    for cc in range(1, last_col + 1):
                        sep_cell = ws.cell(row=row_idx, column=cc, value="")
                        sep_cell.fill = copy(section_fill)
                        sep_cell.border = Border(top=Side(style="thin", color="B8CCE4"))
                    ws.row_dimensions[row_idx].height = 7.5
                    row_idx += 1

        for metric_label, number_format, value_scale in quarterly_metric_order:
            seg_values = quarterly_metrics.get(metric_label, {})
            if seg_values:
                _write_quarterly_segment_metric(
                    metric_label,
                    seg_values,
                    number_format=number_format,
                    value_scale=value_scale,
                )

    annual_segment_data = _load_latest_annual_segment_data()
    row_idx += 1
    annual_section_row = row_idx
    _write_section("Annual segments")
    _promote_section_band(annual_section_row)

    annual_metrics = dict(annual_segment_data.get("metrics") or {})
    annual_assets = dict(annual_segment_data.get("assets") or {})
    available_years = sorted({int(y) for y in annual_segment_data.get("years", []) if str(y).isdigit()})
    year_cols = [yy for yy in available_years if yy in {2023, 2024, 2025}]
    if not year_cols:
        year_cols = available_years[-3:]

    if not annual_metrics and not annual_assets:
        ws[f"A{row_idx}"] = f"No annual segment data found for {str(ticker or 'this ticker').upper()}."
        ws[f"A{row_idx}"].font = Font(italic=True, size=font_size)
        row_idx += 1
    else:
        ws[f"A{row_idx}"] = "Year"
        if valuation_quarter_style_a is not None:
            ws[f"A{row_idx}"]._style = copy(valuation_quarter_style_a)
        else:
            ws[f"A{row_idx}"].font = bold
            ws[f"A{row_idx}"].fill = header_fill
            ws[f"A{row_idx}"].border = thin_border
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
        for idx_y, yy in enumerate(year_cols):
            cc = start_col + idx_y
            cell = ws.cell(row=row_idx, column=cc, value=int(yy))
            if valuation_quarter_style_col is not None:
                cell._style = copy(valuation_quarter_style_col)
            else:
                cell.font = bold
                cell.alignment = Alignment(horizontal="center")
                cell.fill = header_fill
                cell.border = thin_border
        ws.row_dimensions[row_idx].height = 19.5
        row_idx += 1

        default_annual_segments = list(annual_segment_labels) or [
            "Ethanol production",
            "Agribusiness and energy services",
            "Corporate activities",
            "Corporate expense",
            "Corporate assets",
            "Other operations",
            "Intersegment eliminations",
        ]
        annual_metric_order = [
            ("Revenues", default_annual_segments),
            ("Gross margin", default_annual_segments),
            ("Depreciation & amortization", default_annual_segments),
            ("Operating income (loss)", default_annual_segments),
            ("Total assets", default_annual_segments),
        ]

        annual_source_note = str(annual_segment_data.get("source_doc") or "").strip()

        def _write_annual_segment_metric(metric_label: str, seg_values: Dict[str, Dict[int, float]]) -> None:
            nonlocal row_idx
            _write_section(metric_label)
            segment_order = next((order for met, order in annual_metric_order if met == metric_label), list(seg_values.keys()))
            seen_segments: List[str] = []

            def _has_visible_annual_segment_value(segment_label: str) -> bool:
                segment_values = dict(seg_values.get(segment_label, {}) or {})
                for yy in year_cols:
                    raw_val = segment_values.get(int(yy))
                    raw_num = pd.to_numeric(raw_val, errors="coerce")
                    if pd.notna(raw_num):
                        return True
                return False

            for seg in segment_order + [s for s in seg_values.keys() if s not in segment_order]:
                if seg in seen_segments or seg not in seg_values:
                    continue
                if not _has_visible_annual_segment_value(seg):
                    continue
                seen_segments.append(seg)
                ws[f"A{row_idx}"] = seg
                if valuation_label_style is not None:
                    ws[f"A{row_idx}"]._style = copy(valuation_label_style)
                    _apply_data_font(ws[f"A{row_idx}"], bold_flag=False)
                else:
                    ws[f"A{row_idx}"].font = Font(bold=False, size=valuation_data_font_size)
                for idx_y, yy in enumerate(year_cols):
                    cc = start_col + idx_y
                    vv = seg_values.get(seg, {}).get(int(yy))
                    cell_value = None if vv is None else float(vv) / 1e6
                    if is_anf_profile and cell_value is not None:
                        cell_value = round(float(cell_value), 1)
                    cell = ws.cell(row=row_idx, column=cc, value=cell_value)
                    if valuation_numeric_style is not None:
                        cell._style = copy(valuation_numeric_style)
                    _apply_data_font(cell, bold_flag=False)
                    cell.number_format = "#,##0.0"
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="right")
                    if annual_source_note and idx_y == 0:
                        _set_comment(cell, annual_source_note)
                row_hidden_source_values[row_idx] = {
                    int(src_year): float(src_val) / 1e6
                    for src_year, src_val in dict(seg_values.get(seg, {}) or {}).items()
                    if str(src_year).isdigit() and src_val is not None and pd.notna(pd.to_numeric(src_val, errors="coerce"))
                }
                row_visible_keys[row_idx] = [int(yy) for yy in year_cols]
                numeric_rows.append(row_idx)
                row_color_contexts[row_idx] = ("Annual segments", metric_label)
                row_idx += 1

        for metric_label, _segment_order in annual_metric_order:
            if metric_label == "Total assets":
                seg_values = annual_assets
            else:
                seg_values = annual_metrics.get(metric_label, {})
            if seg_values:
                _write_annual_segment_metric(metric_label, seg_values)

    major_band_labels = {"Balance sheet & Segments", "Quarterly segments", "Annual segments"}
    subheader_labels = {"Actuals", "Quarter", "Year"}
    for rr in range(8, ws.max_row + 1):
        label_txt = str(ws.cell(row=rr, column=1).value or "").strip()
        row_vals = [str(ws.cell(row=rr, column=cc).value or "").strip() for cc in range(2, last_col + 1)]
        if label_txt in major_band_labels:
            ws.row_dimensions[rr].height = 24.0
            continue
        if label_txt in subheader_labels:
            ws.row_dimensions[rr].height = 19.5
            continue
        if any(re.fullmatch(r"\d{4}(?:-Q[1-4])?", rv) for rv in row_vals if rv):
            ws.row_dimensions[rr].height = 19.5
            continue
        if rr >= 12:
            ws.row_dimensions[rr].height = 18.0

    # BS + segment QA (latest quarter)
    latest_q = pd.Timestamp(qs[-1]) if qs else None
    qa_status_bits: List[str] = []
    if latest_q is not None:
        # Surface missing key BS sub-lines clearly for latest quarter.
        for key_metric, key_label in [
            ("ap", "Accounts payable"),
            ("accrued", "Accrued liabilities"),
            ("debt_st", "Short-term debt (current portion)"),
            ("ppne", "Property, plant & equipment (net)"),
        ]:
            key_val = maps.get(key_metric, {}).get(latest_q)
            if key_val is None:
                status = "warn"
                source = "History_Q"
                message = f"{key_label} missing for latest quarter (no mapped fact for {latest_q.date()})."
                issue_family = "source_coverage_gap"
                if key_metric == "ap":
                    combined_ap_accrued = maps.get("accrued", {}).get(latest_q)
                    if combined_ap_accrued is not None:
                        status = "info"
                        source = "History_Q/XBRL combined AP+accrued"
                        message = (
                            f"Accounts payable is not separately mapped for {latest_q.date()}; "
                            f"combined accounts payable + accrued liabilities is available at "
                            f"{float(combined_ap_accrued) / 1_000_000.0:,.1f}m. Treat as a source coverage "
                            "limitation, not a parser conflict."
                        )
                qa_rows.append(
                    {
                        "quarter": latest_q.date(),
                        "metric": "BS_Segments",
                        "check": f"{key_metric}_availability",
                        "status": status,
                        "message": message,
                        "source": source,
                        "issue_family": issue_family,
                    }
                )
        a = maps["assets"].get(latest_q)
        l = maps["liabilities"].get(latest_q)
        e = maps["equity"].get(latest_q)
        if a is not None and l is not None and e is not None:
            diff = abs(float(a) - (float(l) + float(e)))
            st = "pass" if diff <= 5_000_000 else ("warn" if diff <= 20_000_000 else "fail")
            qa_rows.append(
                {
                    "quarter": latest_q.date(),
                    "metric": "BS_Segments",
                    "check": "assets_liab_equity_tie_out",
                    "status": st,
                    "message": f"Assets {(a/1e6):,.1f}m vs Liab+Eq {((l+e)/1e6):,.1f}m (diff {(diff/1e6):,.1f}m).",
                    "source": "History_Q",
                }
            )
            qa_status_bits.append(f"A=L+E {st.upper()}")
        st_debt = maps["debt_st"].get(latest_q)
        lt_debt = debt_lt_map.get(latest_q)
        debt_core = maps["debt_total"].get(latest_q)
        if st_debt is not None and lt_debt is not None and debt_core is not None:
            diff_d = abs((float(st_debt) + float(lt_debt)) - float(debt_core))
            st = "pass" if diff_d <= 5_000_000 else ("warn" if diff_d <= 20_000_000 else "fail")
            qa_rows.append(
                {
                    "quarter": latest_q.date(),
                    "metric": "BS_Segments",
                    "check": "debt_tie_out",
                    "status": st,
                    "message": f"ST+LT debt {(float(st_debt)+float(lt_debt))/1e6:,.1f}m vs debt_core {float(debt_core)/1e6:,.1f}m (diff {diff_d/1e6:,.1f}m).",
                    "source": "History_Q",
                }
            )
            qa_status_bits.append(f"Debt {st.upper()}")
        else:
            qa_rows.append(
                {
                    "quarter": latest_q.date(),
                    "metric": "BS_Segments",
                    "check": "debt_tie_out",
                    "status": "warn",
                    "message": "Debt tie-out not attempted (missing short-term debt and/or debt_core components).",
                    "source": "History_Q",
                }
            )
            qa_status_bits.append("Debt N/A")

        cash_latest = maps["cash"].get(latest_q)
        if cash_latest is not None:
            st = "pass" if float(cash_latest) >= 0 else "fail"
            qa_rows.append(
                {
                    "quarter": latest_q.date(),
                    "metric": "BS_Segments",
                    "check": "cash_non_negative",
                    "status": st,
                    "message": f"Cash {float(cash_latest)/1e6:,.1f}m.",
                    "source": "History_Q",
                }
            )
            qa_status_bits.append(f"Cash {st.upper()}")

        quarterly_revenue_values = quarterly_metrics.get("Revenue", {})
        quarterly_source_q = quarterly_segment_data.get("source_qd")
        if quarterly_revenue_values and isinstance(quarterly_source_q, date):
            latest_qd = pd.Timestamp(latest_q).date()
            visible_rev = 0.0
            has_visible_rev = False
            for seg_nm, by_quarter in quarterly_revenue_values.items():
                if "total" in str(seg_nm).strip().lower():
                    continue
                v_seg = by_quarter.get(pd.Timestamp(latest_q))
                if v_seg is None:
                    continue
                visible_rev += float(v_seg)
                has_visible_rev = True
            if has_visible_rev:
                qa_rows.append(
                    {
                        "quarter": latest_qd,
                        "metric": "BS_Segments",
                        "check": "quarterly_segment_data",
                        "status": "pass",
                        "message": f"Quarterly segment block populated for {latest_qd} (visible revenue {visible_rev/1e6:,.1f}m).",
                        "source": str(quarterly_segment_data.get("source_doc") or "segment_financials"),
                    }
                )
                qa_status_bits.append("Quarterly Seg PASS")
            else:
                qa_status_bits.append("Quarterly Seg N/A")
        else:
            qa_status_bits.append("Quarterly Seg N/A")

        annual_revenue_values = annual_metrics.get("Revenues", {})
        fy_source_q = annual_segment_data.get("source_qd")
        if annual_revenue_values and isinstance(fy_source_q, date):
            latest_fy = _annual_segment_latest_year_for_qa(
                annual_revenue_values,
                fy_source_q,
                is_anf_profile=is_anf_profile,
            )
            if latest_fy is None:
                qa_status_bits.append("Annual Seg N/A")
                latest_fy = int(fy_source_q.year)
            seg_sum = 0.0
            has_any_seg = False
            for seg_nm, by_year in annual_revenue_values.items():
                if "intersegment eliminations" in str(seg_nm).strip().lower():
                    continue
                v_seg = by_year.get(latest_fy)
                if v_seg is None:
                    continue
                seg_sum += float(v_seg)
                has_any_seg = True
            elim_v = annual_revenue_values.get("Intersegment eliminations", {}).get(latest_fy)
            total_rev = seg_sum + float(elim_v or 0.0) if has_any_seg else None
            if total_rev is not None:
                qa_rows.append(
                    {
                        "quarter": latest_q.date(),
                        "metric": "BS_Segments",
                        "check": "annual_segment_data",
                        "status": "pass",
                        "message": f"Annual segment block populated from FY{latest_fy} 10-K (revenue total {total_rev/1e6:,.1f}m).",
                        "source": str(annual_segment_data.get("source_doc") or "financial_statement"),
                    }
                )
                qa_status_bits.append("Annual Seg PASS")
            else:
                qa_status_bits.append("Annual Seg N/A")
        else:
            qa_status_bits.append("Annual Seg N/A")

    ws["A4"] = "QA: " + (" | ".join(qa_status_bits) if qa_status_bits else "No latest-quarter checks")
    ws["A4"].font = Font(bold=True, size=11)

    def _bucket_fill(v: float) -> PatternFill:
        x = float(v)
        if x <= -0.15:
            return copy(fill_neg_strong)
        if x <= -0.05:
            return copy(fill_neg_mild)
        if x < 0.05:
            return copy(fill_neutral)
        if x < 0.15:
            return copy(fill_pos_mild)
        return copy(fill_pos_strong)

    def _solid_bucket_fill(fill_in: Any, fallback_color: str) -> PatternFill:
        try:
            rgb = str(getattr(getattr(fill_in, "fgColor", None), "rgb", "") or "").strip()
        except Exception:
            rgb = ""
        if not rgb or rgb.lower() == "00000000":
            rgb = fallback_color
        return PatternFill("solid", fgColor=rgb)

    fill_neg_strong = _solid_bucket_fill(valuation_bucket_fills.get("neg_strong"), "00A63A00")
    fill_neg_mild = _solid_bucket_fill(valuation_bucket_fills.get("neg_mild"), "00D55E00")
    fill_neutral = _solid_bucket_fill(valuation_bucket_fills.get("neutral"), "00DDDDDD")
    fill_pos_mild = _solid_bucket_fill(valuation_bucket_fills.get("pos_mild"), "009BD3F5")
    fill_pos_strong = _solid_bucket_fill(valuation_bucket_fills.get("pos_strong"), "002F80ED")

    # Give visible value blocks the same restrained valuation-style language before bucket fills.
    for rr in numeric_rows:
        label_cell = ws.cell(row=rr, column=1)
        row_section_label, row_subsection_label = row_color_contexts.get(rr, ("", ""))
        row_policy = _quarterly_row_color_policy(
            label_cell.value,
            section_label=row_section_label,
            subsection_label=row_subsection_label,
        )
        if valuation_label_style is not None and label_cell.value not in (None, ""):
            label_cell._style = copy(valuation_label_style)
            _apply_data_font(label_cell, bold_flag=False)
        for cc in range(start_col, last_col + 1):
            data_cell = ws.cell(row=rr, column=cc)
            if data_cell.value in (None, ""):
                continue
            data_cell.border = thin_border
            data_cell.alignment = Alignment(horizontal="right")
            if (
                row_section_label in {"Quarterly segments", "Annual segments"}
                or rr in delta_nwc_rows
                or rr in nwc_rows
                or rr in no_bucket_rows
                or rr in pct_rows
                or row_policy.directionality == "neutral"
            ):
                data_cell.fill = copy(fill_neutral)

    # Clear sheet-level CF and apply direct bucket fills (Valuation-like visual behavior).
    try:
        ws.conditional_formatting._cf_rules.clear()  # type: ignore[attr-defined]
    except Exception:
        pass

    # Apply direct fills to numeric rows using the shared metric-aware policy.
    if numeric_rows:
        for rr in numeric_rows:
            if rr in delta_nwc_rows or rr in nwc_rows:
                continue
            row_label = ws.cell(row=rr, column=1).value
            row_section_label, row_subsection_label = row_color_contexts.get(rr, ("", ""))
            row_policy = _quarterly_row_color_policy(
                row_label,
                section_label=row_section_label,
                subsection_label=row_subsection_label,
            )
            if row_policy.directionality == "neutral":
                continue
            row_values = [ws.cell(row=rr, column=cc).value for cc in range(start_col, last_col + 1)]
            for idx_cc, cc in enumerate(range(start_col, last_col + 1)):
                cur_cell = ws.cell(row=rr, column=cc)
                metric = _quarterly_color_metric_from_series(
                    row_values,
                    idx_cc,
                    comparison_basis=row_policy.comparison_basis,
                    directionality=row_policy.directionality,
                )
                if metric is None and idx_cc < len(row_visible_keys.get(rr) or []):
                    metric = _hidden_source_comparison_metric(
                        current_key=(row_visible_keys.get(rr) or [])[idx_cc],
                        current_value=cur_cell.value,
                        visible_idx=idx_cc,
                        comparison_basis=row_policy.comparison_basis,
                        directionality=row_policy.directionality,
                        source_values=row_hidden_source_values.get(rr),
                    )
                if metric is None:
                    continue
                cur_cell.fill = _bucket_fill(metric)

    # Delta NWC row uses robust denominator.
    if delta_nwc_rows:
        qts = [pd.Timestamp(x) for x in qs]
        for rr in delta_nwc_rows:
            for i in range(1, len(qts)):
                q_prev = qts[i - 1]
                q_cur = qts[i]
                cc = start_col + i
                cur_cell = ws.cell(row=rr, column=cc)
                delta_raw = pd.to_numeric(cur_cell.value, errors="coerce")
                if pd.isna(delta_raw):
                    continue
                nwc_prev = nwc_map.get(q_prev)
                denom = None
                if nwc_prev is not None and abs(float(nwc_prev)) >= 1_000_000.0:
                    denom = float(nwc_prev)
                else:
                    ca_prev = maps["assets_current"].get(q_prev)
                    if ca_prev is not None and abs(float(ca_prev)) >= 1_000_000.0:
                        denom = float(ca_prev)
                if denom is None or abs(denom) < 1_000_000.0:
                    continue
                ratio = float(delta_raw * 1e6) / abs(float(denom))
                cur_cell.fill = _bucket_fill(ratio)

    ws.freeze_panes = "B12"
    w_a = valuation_col_widths.get("A")
    ws.column_dimensions["A"].width = max(54, float(w_a) if w_a is not None else 54)
    for i in range(start_col, last_col + 1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 12
    # Keep mid quarter columns compact to avoid oversized visual gaps in BS_Segments.
    for letter, idx_letter in (("F", 6), ("G", 7)):
        if start_col <= idx_letter <= last_col:
            ws.column_dimensions[letter].width = 10
    for rr in range(10, ws.max_row + 1):
        for cc in range(1, last_col + 1):
            ws.cell(row=rr, column=cc).border = thin_border
    for r in qa_rows:
        ui_info_rows.append(
            {
                "quarter": r.get("quarter"),
                "metric": "BS_Segments",
                "severity": str(r.get("status") or "info"),
                "message": f"{r.get('check')}: {r.get('message')}",
                "source": r.get("source") or "",
            }
        )
    return qa_rows
