"""Quarterly comparison-fill helpers for workbook rendering."""
from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl.styles import PatternFill


@dataclass(frozen=True)
class QuarterlyRowColorPolicy:
    comparison_basis: str
    directionality: str


_QUARTERLY_COLOR_BRACKET_SUFFIX_RE = re.compile(r"\[[^\]]+\]")
_QUARTERLY_COLOR_REPEAT_SPACE_RE = re.compile(r"\s+")
_QUARTERLY_COLOR_COMPARISON_RE = re.compile(r"\b(?:yoy|qoq)\b", re.I)


def _normalize_quarterly_color_label(label: Any) -> str:
    text = str(label or "").strip()
    if not text:
        return ""
    text = text.replace("Î”", "Δ").replace("\u2212", "-")
    text = _QUARTERLY_COLOR_BRACKET_SUFFIX_RE.sub("", text)
    text = _QUARTERLY_COLOR_REPEAT_SPACE_RE.sub(" ", text)
    return text.strip()


def _quarterly_color_label_key(label: Any) -> str:
    text = _normalize_quarterly_color_label(label)
    if not text:
        return ""
    text = text.replace("Δ", " delta ").replace("δ", " delta ").lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9%]+", " ", text)
    text = _QUARTERLY_COLOR_REPEAT_SPACE_RE.sub(" ", text)
    return text.strip()


def _quarterly_color_basis_for_label(label: Any) -> str:
    raw = _normalize_quarterly_color_label(label)
    key = _quarterly_color_label_key(label)
    if not raw or not key:
        return "yoy"
    if re.search(r"\b(comp|comps|comparable sales)\b", key):
        return "direct_pct_points"
    if "yoy" in key and any(
        token in key
        for token in (
            "sales yoy",
            "net sales yoy",
            "revenue yoy",
            "margin yoy",
            "growth yoy",
        )
    ):
        return "direct_pct_points"
    if _QUARTERLY_COLOR_COMPARISON_RE.search(key) and ("delta" in key or "%" in raw):
        return "direct_delta"
    # Capital-return execution rows are quarter-native flows. QoQ is the
    # meaningful comparator and avoids leaving live quarter buyback rows
    # visually blank just because there is no year-ago execution baseline.
    if key == "buybacks cash":
        return "qoq"
    if "qoq" in key:
        return "qoq"
    if "ttm" in key:
        return "ttm_vs_prior_ttm"
    return "yoy"


def _quarterly_color_directionality_for_label(
    label: Any,
    *,
    section_label: str = "",
    subsection_label: str = "",
) -> str:
    key = _quarterly_color_label_key(label)
    section_key = _quarterly_color_label_key(section_label)
    subsection_key = _quarterly_color_label_key(subsection_label)
    if not key:
        return "neutral"

    segment_result_metric_keys = {
        "revenue",
        "revenues",
        "gross margin",
        "adjusted ebit",
        "operating income loss",
        "ebit margin %",
    }
    segment_neutral_metric_keys = {
        "depreciation amortization",
        "total assets",
    }
    segment_neutral_label_keys = {
        "intersegment eliminations",
        "corporate assets",
    }
    segment_profit_only_label_keys = {
        "corporate expense",
        "corporate activities",
    }

    if key == "acquisitions ttm cash":
        return "neutral"

    if subsection_key == "per share earnings":
        return "higher_better"

    if section_key == "operating":
        return "higher_better"

    if section_key in {"quarterly segments", "annual segments"}:
        if key in segment_neutral_label_keys:
            return "neutral"
        if subsection_key in segment_neutral_metric_keys:
            return "neutral"
        if key in segment_profit_only_label_keys:
            if subsection_key in {"gross margin", "adjusted ebit", "operating income loss", "ebit margin %"}:
                return "higher_better"
            return "neutral"
        if subsection_key in segment_result_metric_keys:
            return "higher_better"

    if key == "interest coverage" or key.startswith("interest coverage "):
        return "higher_better"

    if key == "cash interest coverage" or key.startswith("cash interest coverage "):
        return "higher_better"

    if key in {
        "cfo",
        "fcf cfo capex",
        "fcf yoy delta m",
        "fcf ttm yoy delta m",
        "fcf ttm yoy m",
        "fcf ttm",
        "adj fcf ttm",
        "adj fcf fcf",
        "owner earnings proxy",
        "fcf margin %",
        "fcf margin ttm",
        "buybacks cash",
        "buybacks ttm cash",
        "dividends ttm cash",
        "debt repaid gross ttm",
        "cash",
        "cash and cash equivalents",
        "restricted cash",
        "total cash restricted cash",
        "short term investments",
        "net working capital",
        "current ratio",
        "quick ratio",
        "total equity",
        "finance receivables total",
        "deposits bank customer",
        "interest coverage p and l ttm",
        "cash interest coverage ttm",
        "ebitda ttm",
        "adj ebitda ttm",
        "fcf conversion ttm",
        "bv share",
        "tbv share",
        "fcf share ttm",
        "revolver availability",
        "revolver capacity",
        "liquidity cash availability",
    }:
        return "higher_better"

    if "sales yoy" in key or key.endswith(" yoy") and any(tok in key for tok in ("sales", "margin", "growth")):
        return "higher_better"

    if any(
        token in key
        for token in (
            "ethanol gallons sold",
            "ethanol gallons produced",
            "ultra high protein",
            "renewable corn oil",
            "distillers grains",
        )
    ):
        return "higher_better"

    if key in {
        "capex",
        "capex % of revenue",
        "capex % of revenue ttm",
        "interest paid",
        "tax paid",
        "goodwill % of assets",
        "accounts payable",
        "accrued liabilities",
        "derivative financial instruments liability",
        "short term notes payable and other borrowings",
        "current maturities of long term debt",
        "operating lease current liabilities",
        "total current liabilities",
        "debt issued gross ttm",
        "long term debt",
        "carbon equipment liabilities",
        "operating lease long term liabilities",
        "other long term liabilities",
        "total liabilities",
        "bank net funding",
        "debt core",
        "debt core borrowings",
        "net pension opeb",
        "net debt core",
        "net debt core borrowings",
        "net debt qoq delta m",
        "net debt yoy delta m",
        "net leverage",
        "net leverage adj",
        "diluted shares m",
        "shares diluted m",
        "shares outstanding m",
        "revolver drawn",
        "revolver letters of credit",
    }:
        return "lower_better"

    if key.startswith("shares qoq delta") or key.startswith("shares yoy delta"):
        return "lower_better"

    if "pension" in key and "opeb" in key:
        return "lower_better"

    if "total debt" in key and ("delta" in key or "qoq" in key or "yoy" in key):
        return "lower_better"

    if "cash" in key and ("delta" in key or "qoq" in key or "yoy" in key):
        return "higher_better"

    return "neutral"


def _quarterly_row_color_policy(
    label: Any,
    *,
    section_label: str = "",
    subsection_label: str = "",
) -> QuarterlyRowColorPolicy:
    return QuarterlyRowColorPolicy(
        comparison_basis=_quarterly_color_basis_for_label(label),
        directionality=_quarterly_color_directionality_for_label(
            label,
            section_label=section_label,
            subsection_label=subsection_label,
        ),
    )


def _quarterly_color_metric_from_series(
    row_values: List[Any],
    idx: int,
    *,
    comparison_basis: str,
    directionality: str,
) -> Optional[float]:
    if directionality not in {"higher_better", "lower_better"}:
        return None
    if idx < 0 or idx >= len(row_values):
        return None
    current = pd.to_numeric(row_values[idx], errors="coerce")
    if pd.isna(current):
        return None
    if comparison_basis in {"direct_delta", "direct_pct_points"}:
        metric = float(current)
        if comparison_basis == "direct_pct_points" and abs(metric) > 1.0:
            metric /= 100.0
    else:
        step = 1 if comparison_basis == "qoq" else 4
        prev_idx = idx - step
        if prev_idx < 0:
            return None
        previous = pd.to_numeric(row_values[prev_idx], errors="coerce")
        if pd.isna(previous) or abs(float(previous)) <= 1e-12:
            return None
        metric = (float(current) - float(previous)) / abs(float(previous))
    if directionality == "lower_better":
        metric *= -1.0
    return metric


def _quarterly_bucket_fill(v: Any) -> Optional[PatternFill]:
    num = pd.to_numeric(v, errors="coerce")
    if pd.isna(num):
        return None
    x = float(num)
    if x <= -0.15:
        return PatternFill("solid", fgColor="A63A00")
    if x <= -0.05:
        return PatternFill("solid", fgColor="D55E00")
    if x <= 0.05:
        return PatternFill("solid", fgColor="DDDDDD")
    if x <= 0.15:
        return PatternFill("solid", fgColor="9BD3F5")
    return PatternFill("solid", fgColor="2F80ED")


def _hidden_source_comparison_metric(
    *,
    current_key: Any,
    current_value: Any,
    visible_idx: int,
    comparison_basis: str,
    directionality: str,
    source_values: Optional[Dict[Any, Any]],
) -> Optional[float]:
    if directionality not in {"higher_better", "lower_better"} or comparison_basis == "direct_delta" or not source_values:
        return None
    current_num = pd.to_numeric(current_value, errors="coerce")
    if pd.isna(current_num):
        return None

    source_quarter_map: Dict[pd.Timestamp, float] = {}
    source_year_map: Dict[int, float] = {}
    for raw_key, raw_val in dict(source_values or {}).items():
        raw_num = pd.to_numeric(raw_val, errors="coerce")
        if pd.isna(raw_num):
            continue
        raw_key_txt = str(raw_key).strip()
        if re.fullmatch(r"\d{4}", raw_key_txt):
            source_year_map[int(raw_key_txt)] = float(raw_num)
            continue
        raw_ts = pd.to_datetime(raw_key, errors="coerce")
        if pd.isna(raw_ts):
            continue
        source_quarter_map[pd.Timestamp(raw_ts).normalize()] = float(raw_num)

    current_key_txt = str(current_key).strip()
    if re.fullmatch(r"\d{4}", current_key_txt):
        if visible_idx >= 1:
            return None
        previous = pd.to_numeric(source_year_map.get(int(current_key_txt) - 1), errors="coerce")
        if pd.isna(previous) or abs(float(previous)) <= 1e-12:
            return None
        metric = (float(current_num) - float(previous)) / abs(float(previous))
        if directionality == "lower_better":
            metric *= -1.0
        return metric

    current_ts = pd.to_datetime(current_key, errors="coerce")
    if pd.isna(current_ts):
        return None
    step = 1 if comparison_basis == "qoq" else 4
    if visible_idx >= step:
        return None
    try:
        prev_period = pd.Timestamp(current_ts).to_period("Q") - step
        prev_q = prev_period.end_time.normalize()
    except Exception:
        return None

    def _ordered_previous_value() -> Optional[float]:
        current_norm = pd.Timestamp(current_ts).normalize()
        source_keys = sorted(source_quarter_map)
        try:
            current_pos = source_keys.index(current_norm)
            matched_current = current_norm
        except ValueError:
            nearest = sorted(
                (
                    (abs((current_norm - source_key).days), idx_key, source_key)
                    for idx_key, source_key in enumerate(source_keys)
                    if abs((current_norm - source_key).days) <= 45
                ),
                key=lambda item: item[0],
            )
            if not nearest:
                return None
            _, current_pos, matched_current = nearest[0]
        if current_pos < step:
            return None
        candidate_key = source_keys[current_pos - step]
        days_delta = abs((matched_current - candidate_key).days)
        if step == 1:
            if not 45 <= days_delta <= 125:
                return None
        else:
            if not 330 <= days_delta <= 400:
                return None
        candidate = pd.to_numeric(source_quarter_map.get(candidate_key), errors="coerce")
        if pd.isna(candidate):
            return None
        return float(candidate)

    previous = pd.to_numeric(source_quarter_map.get(prev_q), errors="coerce")
    if pd.isna(previous):
        previous = pd.to_numeric(_ordered_previous_value(), errors="coerce")
    if pd.isna(previous) or abs(float(previous)) <= 1e-12:
        return None
    metric = (float(current_num) - float(previous)) / abs(float(previous))
    if directionality == "lower_better":
        metric *= -1.0
    return metric


def _apply_quarterly_comparison_fills(
    cells: List[Any],
    row_values: List[Any],
    *,
    label: Any,
    section_label: str = "",
    subsection_label: str = "",
    visible_keys: Optional[List[Any]] = None,
    source_values: Optional[Dict[Any, Any]] = None,
) -> None:
    if not cells or not row_values:
        return
    policy = _quarterly_row_color_policy(
        label,
        section_label=section_label,
        subsection_label=subsection_label,
    )
    for idx, cell in enumerate(cells):
        metric = _quarterly_color_metric_from_series(
            row_values,
            idx,
            comparison_basis=policy.comparison_basis,
            directionality=policy.directionality,
        )
        if metric is None and visible_keys is not None and idx < len(visible_keys):
            metric = _hidden_source_comparison_metric(
                current_key=visible_keys[idx],
                current_value=row_values[idx],
                visible_idx=idx,
                comparison_basis=policy.comparison_basis,
                directionality=policy.directionality,
                source_values=source_values,
            )
        fill = _quarterly_bucket_fill(metric)
        if fill is not None:
            cell.fill = fill
