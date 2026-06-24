"""Main workbook rendering context and visible-sheet write logic.

This module is the largest concentration of product behavior in the codebase. It turns
pipeline artifacts, local materials, SEC cache content, and market-data exports into
the actual saved workbook surfaces delivered to the user.

When debugging workbook output, this is usually the last stop before a value becomes
visible in the saved file.
"""
from __future__ import annotations

import datetime as dt
import hashlib
import html
import io
import json
import math
import os
import re
import time
from contextlib import contextmanager
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Mapping, Optional, Pattern, Sequence, Set, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines, TextAxis
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties, LineProperties
from openpyxl.comments import Comment
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.datetime import to_excel
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from .conference_metadata import metadata_audit_flags, metadata_source_file, parse_metadata_key_values, source_material_role

from .excel_writer_bs_segments import (
    BsSegmentsWriterDeps,
    _series_has_nonblank_values,
    _should_render_carbon_equipment_liabilities,
    write_bs_segments_sheet,
)
from .excel_writer_bs_segments_sheet_adapter import (
    BSSegmentsSheetAdapter,
    BSSegmentsSheetAdapterDeps,
)

from .excel_writer_coloring import (
    QuarterlyRowColorPolicy,
    _QUARTERLY_COLOR_BRACKET_SUFFIX_RE,
    _QUARTERLY_COLOR_COMPARISON_RE,
    _QUARTERLY_COLOR_REPEAT_SPACE_RE,
    _apply_quarterly_comparison_fills,
    _hidden_source_comparison_metric,
    _normalize_quarterly_color_label,
    _quarterly_bucket_fill,
    _quarterly_color_basis_for_label,
    _quarterly_color_directionality_for_label,
    _quarterly_color_label_key,
    _quarterly_color_metric_from_series,
    _quarterly_row_color_policy,
)
from .excel_writer_layout import (
    estimate_wrapped_line_count as _estimate_wrapped_line_count,
    estimate_wrapped_row_height as _estimate_wrapped_row_height,
)
from .excel_writer_analysis_sheet_layout_support import (
    AnalysisSheetLayoutSupport,
    AnalysisSheetLayoutSupportDeps,
)
from .excel_writer_chart_text_support import (
    ChartTextSupport,
    ChartTextSupportDeps,
)
from .excel_writer_evidence_source_support import (
    EvidenceSourceSupport,
    EvidenceSourceSupportDeps,
)
from .excel_writer_sector_operating_driver_intro_support import (
    SectorOperatingDriverIntroSupport,
    SectorOperatingDriverIntroSupportDeps,
)
from .excel_writer_sector_investment_case_support import (
    SectorInvestmentCaseSupport,
    SectorInvestmentCaseSupportDeps,
)
from .excel_writer_investment_case_readability import (
    InvestmentCaseReadability,
    InvestmentCaseReadabilityDeps,
)
from .excel_writer_operating_drivers import (
    OperatingDriversWriterDeps,
    write_operating_drivers_sheet,
)
from .excel_writer_operating_drivers_sheet_adapter import (
    OperatingDriversSheetAdapter,
    OperatingDriversSheetAdapterDeps,
)
from .excel_writer_operating_drivers_raw_sheet import (
    OperatingDriversRawSheetDeps,
    OperatingDriversRawSheetWriter,
)
from .excel_writer_economics_overlay_charts import (
    EconomicsOverlayChartWriterDeps,
    _build_visible_quarter_label_points,
    _quarter_bounds_from_end_date,
    write_economics_overlay_charts,
)
from .excel_writer_economics_overlay_commercial import (
    GpreEconomicsOverlayCommercialDeps,
    write_gpre_economics_overlay_commercial_sections,
)
from .excel_writer_economics_overlay_bridge import (
    GpreEconomicsOverlayBridgeDeps,
    write_gpre_economics_overlay_bridge_to_reported_section,
)
from .excel_writer_economics_overlay_current_qtd import (
    GpreEconomicsOverlayCurrentQtdDeps,
    write_gpre_economics_overlay_current_qtd_section,
)
from .excel_writer_economics_overlay_inputs import (
    GpreEconomicsOverlayInputRowsDeps,
    write_gpre_economics_overlay_input_rows,
)
from .excel_writer_economics_overlay_coproduct import (
    GpreEconomicsOverlayCoproductDeps,
    write_gpre_economics_overlay_coproduct_section,
)
from .excel_writer_economics_overlay_sources import (
    EconomicsOverlaySourceSupport,
    EconomicsOverlaySourceSupportDeps,
)
from .excel_writer_economics_overlay_market_state import (
    EconomicsOverlayMarketStateDeps,
    build_economics_overlay_market_state,
)
from .excel_writer_economics_overlay_sheet import (
    EconomicsOverlaySheetDeps,
    EconomicsOverlaySheetWriter,
)
from .excel_writer_economics_raw import (
    ECONOMICS_MARKET_RAW_COLUMN_WIDTHS,
    ECONOMICS_MARKET_RAW_HEADERS,
    ECONOMICS_MARKET_RAW_LARGE_FAST_PATH_THRESHOLD,
    EconomicsMarketRawWriterDeps,
    write_economics_market_raw_sheet,
)
from .excel_writer_market_data_sources import (
    EconomicsMarketRowsDeps,
    _convert_market_price_value,
    _economics_market_region_tags,
    _economics_market_series_meta,
    build_economics_market_rows,
)
from .excel_writer_styles import (
    get_analysis_sheet_style_bundle as _style_get_analysis_sheet_style_bundle,
    valuation_side_panel_style_bundle as _style_valuation_side_panel_style_bundle,
)
from .excel_writer_valuation import (
    build_valuation_history_source_maps,
    display_m_source_map,
    history_margin_source_map,
    history_numeric_source_map,
    normalize_capex_for_valuation,
    quarter_key_union,
    ttm_map,
    ttm_sparse_cashflow_map,
    valuation_hidden_comparison_metric,
)
from .excel_writer_valuation_precompute import (
    ValuationPrecomputeDeps,
    ValuationPrecomputeSupport,
)
from .excel_writer_valuation_render_bundle import (
    ValuationRenderBundleDeps,
    ensure_valuation_render_bundle,
)
from .excel_writer_valuation_style_bundle import (
    ValuationStyleBundleDeps,
    get_valuation_style_bundle,
)
from .excel_writer_valuation_debt_support import (
    ValuationDebtSupportDeps,
    source_backed_debt_tranches_from_slides,
)
from .excel_writer_valuation_bridge_support import (
    ValuationBridgeSupport,
    ValuationBridgeSupportDeps,
)
from .excel_writer_local_balance_sheet_support import (
    LocalBalanceSheetSupport,
    LocalBalanceSheetSupportDeps,
)
from .excel_writer_history_q_fiscal_support import (
    HistoryQFiscalSupport,
    HistoryQFiscalSupportDeps,
)
from .excel_writer_valuation_orchestrator import (
    ValuationOrchestratorDeps,
    write_valuation_sheet,
)
from .excel_writer_summary_builder import (
    SummaryBuilderDeps,
    build_summary_dataframe,
)
from .excel_writer_latest_quarter_qa import (
    LatestQuarterQADeps,
    LatestQuarterQASupport,
    run_latest_quarter_qa,
)
from .excel_writer_gpre_commercial_setup import (
    GpreCommercialSetupDeps,
    GpreCommercialSetupSupport,
)
from .excel_writer_operating_drivers_support import (
    OperatingDriversSupport,
    OperatingDriversSupportDeps,
)
from .excel_writer_operating_driver_workbook_support import (
    OperatingDriverWorkbookSupport,
    OperatingDriverWorkbookSupportDeps,
)
from .excel_writer_profile_signal_support import (
    ProfileSignalSupport,
    ProfileSignalSupportDeps,
)
from .excel_writer_promise_source_override_mutator import (
    PromiseSourceOverrideMutatorDeps,
    apply_source_backed_promise_mapping_overrides,
)
from .excel_writer_promise_progress_rewrite import (
    PromiseProgressRewriteDeps,
    rewrite_shared_promise_progress_ui_from_blocks,
)
from .excel_writer_promise_progress_worksheet_repairs import (
    PROMISE_TIMELINE_HEADERS,
    PROMISE_VISIBLE_MAX_COL,
    PromiseProgressWorksheetRepairDeps,
    final_repair_promise_progress_ui,
    insert_management_credibility_scorecard,
    polish_promise_scorecard_layout,
    repair_promise_table_header_merges,
)
from . import excel_writer_promise_progress_worksheet_repairs as _promise_progress_worksheet_repairs
from .excel_writer_shared_ui_conventions import (
    SharedUiConventionsDeps,
    apply_shared_ui_conventions_to_workbook,
)
from .excel_writer_anf_valuation_side_panel import (
    AnfValuationSidePanelDeps,
    clear_anf_valuation_side_panels,
    valuation_side_panel_style_bundle,
    write_anf_valuation_side_panel,
)
from .excel_writer_anf_valuation_support import (
    AnfValuationSupport,
    AnfValuationSupportDeps,
)
from .excel_writer_anf_visible_support import (
    AnfVisibleSupport,
    AnfVisibleSupportDeps,
)
from .excel_writer_financial_report_support import (
    FinancialReportSupport,
    FinancialReportSupportDeps,
)
from .excel_writer_investment_case_support import (
    InvestmentCaseSupport,
    InvestmentCaseSupportDeps,
)
from .excel_writer_sector_investment_case import (
    SectorInvestmentCaseRenderDeps,
    write_sector_investment_case_data_sheet,
    write_sector_investment_case_sheet,
)
from .excel_writer_anf_investment_case import (
    AnfInvestmentCaseRenderDeps,
    write_anf_investment_case_data_sheet,
    write_anf_investment_case_sheet,
)
from .excel_writer_investment_case_scenarios import (
    InvestmentCaseScenarioRenderDeps,
    SCENARIO_DRIVER_CAPITAL_STRUCTURE_INTEREST,
    SCENARIO_DRIVER_CASH_FLOW_CAPEX,
    SCENARIO_DRIVER_MANUAL_INCREMENTAL,
    SCENARIO_DRIVER_MARGIN_EBITDA,
    SCENARIO_DRIVER_REVENUE_VOLUME,
    SCENARIO_DRIVER_SHARE_COUNT_BUYBACK,
    SCENARIO_DRIVER_TAX_CREDIT_SUBSIDY,
    SCENARIO_TAX_CASH_ONLY,
    SCENARIO_TAX_DIRECT_EPS,
    SCENARIO_TAX_NON_TAXABLE,
    SCENARIO_TAX_NON_TAXABLE_CREDIT,
    SCENARIO_TAX_NO_EPS_IMPACT,
    SCENARIO_TAX_TAXABLE,
    SCENARIO_TAX_UNKNOWN_MANUAL_REQUIRED,
    _ScenarioDriverBridgeSpec,
    _SegmentScenarioInputSpec,
    _excel_manual_percent_active_formula,
    _excel_percent_value_expr,
    _excel_visible_value_range_formula,
    _scenario_bridge_active_value_formula,
    _scenario_bridge_default_eps_rule,
    _scenario_bridge_eps_manual_required,
    _scenario_bridge_eps_value_formula,
    _scenario_bridge_incremental_formula,
    _scenario_bridge_negative_impact_formula,
    _scenario_bridge_row_values,
    _scenario_bridge_same_impact_formula,
    _scenario_bridge_tax_audit_rows,
    _scenario_bridge_tax_conversion_label,
    _segment_scenario_margin_value,
    _segment_scenario_note_for_view,
    _segment_scenario_revenue_m,
    _segment_scenario_specs_from_records,
    _segment_scenario_view_basis,
    write_scenario_bridge_tax_treatment_sheet as _write_scenario_bridge_tax_treatment_sheet_impl,
    write_scenario_driver_assumptions_sheet as _write_scenario_driver_assumptions_sheet_impl,
)
from .excel_writer_quarter_notes_ui_orchestrator import (
    QuarterNotesUiOrchestratorDeps,
    write_quarter_notes_ui_sheet,
)
from .excel_writer_quarter_notes_context_adapter import (
    QuarterNotesContextAdapterDeps,
    standardize_quarter_notes_ui_categories as _standardize_quarter_notes_ui_categories_impl,
    write_quarter_narrative_data_surface as _write_quarter_narrative_data_surface_impl,
    write_quarter_notes_narrative_ui_surface as _write_quarter_notes_narrative_ui_surface_impl,
    write_quarter_notes_ui_v2 as _write_quarter_notes_ui_v2_impl,
)
from .excel_writer_quarter_narrative import (
    QuarterNarrativeRecord,
    QUARTER_NARRATIVE_DATA_HEADERS,
    _quarter_narrative_record_to_audit_row,
    _quarter_narrative_records_for_ticker,
    _write_quarter_narrative_data_sheet,
    _quarter_narrative_period_sort_key,
    _quarter_narrative_source_label,
    _quarter_narrative_compact_sentence,
    _quarter_narrative_row_height,
    _quarter_narrative_recent_history_periods,
    _quarter_narrative_recent_periods_from_frame,
    _quarter_narrative_clean_text,
    _quarter_narrative_period_from_source_quarter,
    _quarter_narrative_category_theme,
    _quarter_narrative_implications_for_row,
    _quarter_narrative_records_from_quarter_notes,
    _quarter_narrative_period_from_label_or_date,
    _quarter_narrative_format_surface_value,
    _quarter_narrative_amount_from_surface_value,
    _quarter_narrative_source_date_from_period,
    _quarter_narrative_surface_row_terms,
    _quarter_narrative_records_from_history_q,
    _quarter_narrative_records_from_operating_drivers,
    _quarter_narrative_records_from_promise_progress_ui,
    _quarter_narrative_records_from_workbook_surfaces,
    _quarter_narrative_records_for_context,
    _quarter_narrative_read_block,
    _write_quarter_notes_ui_narrative_sheet,
)
from .excel_writer_segment_sources import (
    _annual_segment_latest_year_for_qa,
    _anf_add_total_company_quarter_revenue_from_history,
    _anf_annual_segment_data_from_slides_segments,
    _anf_fill_brand_quarter_revenue_from_annual_segments_for_bs,
    _filter_anf_quarterly_segment_actual_rows,
    _pbi_add_corporate_reconciliation_from_release_text,
    _pbi_repair_total_reportable_segment_quarterly_totals_for_bs,
)


def _guidance_source_contract_label(ticker: Any) -> str:
    """Visible guidance-source contract for shared UI sheets.

    Long-term rule (Option B): Guidance_Normalized is canonical only when it
    contains clean, reliable normalized rows.  If a ticker is still handled by
    curated profile / Slides_Guidance logic, keep Guidance_Normalized empty
    rather than filling it with noisy rows, and do not cite it in visible UI.
    """
    ticker_txt = str(ticker or "").strip().upper()
    if ticker_txt in {"PBI", "GPRE"}:
        return "Slides_Guidance / curated guidance profile"
    return "Guidance_Normalized"


def _gpre_45z_all_facilities_confirmed(*parts: Any) -> bool:
    """Return True when source text confirms all GPRE 45Z facilities/plants."""

    blob = glx_normalize_text(" | ".join(str(part or "") for part in parts)).lower()
    if not blob:
        return False
    has_all = bool(re.search(r"\ball\s+(?:8|eight)\b", blob, re.I))
    has_facility = bool(re.search(r"\b(plants?|facilit(?:y|ies)|operating plants?)\b", blob, re.I))
    has_45z = bool(re.search(r"\b45z\b|production tax credits?|tax credits?", blob, re.I))
    has_confirming_status = bool(
        re.search(
            r"\b(qualified|qualify|qualifying|operational|operating|running|in\s+operation|from\s+jan(?:uary)?\.?\s*1)\b",
            blob,
            re.I,
        )
    )
    return has_all and has_facility and has_45z and has_confirming_status


def _date_is_missing_or_outside(value: Any, start: date, end: date) -> bool:
    try:
        parsed = pd.Timestamp(str(value or "")).date()
    except Exception:
        return True
    return parsed < start or parsed > end





def _company_operating_margin_proxy_from_workbook(wb: Workbook) -> Tuple[Optional[float], str]:
    return SectorInvestmentCaseSupport(
        SectorInvestmentCaseSupportDeps(
            runtime={
                "pd": pd,
                "math": math,
                "re": re,
                "get_column_letter": get_column_letter,
            }
        )
    ).company_operating_margin_proxy_from_workbook(wb)


def _bs_segments_latest_segment_margin_from_workbook(wb: Workbook, label: Any) -> Tuple[Any, str]:
    return SectorInvestmentCaseSupport(
        SectorInvestmentCaseSupportDeps(
            runtime={
                "pd": pd,
                "math": math,
                "re": re,
                "get_column_letter": get_column_letter,
            }
        )
    ).bs_segments_latest_segment_margin_from_workbook(wb, label)


def _date_or_none(value: Any) -> Optional[date]:
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return None
    return pd.Timestamp(ts).date()


def _history_q_fiscal_support() -> HistoryQFiscalSupport:
    return HistoryQFiscalSupport(
        HistoryQFiscalSupportDeps(
            runtime={
                "pd": pd,
                "math": math,
                "re": re,
                "_date_or_none": _date_or_none,
            }
        )
    )


def _fiscal_profile_from_workbook(
    wb: Optional[Workbook],
    ticker: Any = "",
    fiscal_profile: Any = None,
) -> Any:
    return _history_q_fiscal_support().fiscal_profile_from_workbook(
        wb,
        ticker=ticker,
        fiscal_profile=fiscal_profile,
    )


def _history_q_latest_full_year_period_set(
    wb: Workbook,
    *,
    ticker: Any = "",
    fiscal_profile: Any = None,
) -> Dict[str, Any]:
    return _history_q_fiscal_support().history_q_latest_full_year_period_set(
        wb,
        ticker=ticker,
        fiscal_profile=fiscal_profile,
    )


def _history_q_latest_full_year_actuals_from_workbook(
    wb: Workbook,
    *,
    ticker: Any = "",
    fiscal_profile: Any = None,
) -> Dict[str, float]:
    return _history_q_fiscal_support().history_q_latest_full_year_actuals_from_workbook(
        wb,
        ticker=ticker,
        fiscal_profile=fiscal_profile,
    )


def _augment_history_q_frame_for_writer(
    df: pd.DataFrame,
    *,
    ticker: Any = "",
    fiscal_profile: Any = None,
) -> pd.DataFrame:
    return _history_q_fiscal_support().augment_history_q_frame_for_writer(
        df,
        ticker=ticker,
        fiscal_profile=fiscal_profile,
    )


def _history_q_year_default_formulas(
    start_date: Optional[Tuple[int, int, int]] = None,
    end_date: Optional[Tuple[int, int, int]] = None,
    *,
    fiscal_year: Optional[int] = None,
    quarter_dates: Optional[Sequence[date]] = None,
    previous_quarter_dates: Optional[Sequence[date]] = None,
    quarter_criteria: Optional[Sequence[Any]] = None,
    previous_quarter_criteria: Optional[Sequence[Any]] = None,
    start_exclusive: bool = False,
    end_inclusive: bool = False,
) -> Dict[str, str]:
    return _history_q_fiscal_support().history_q_year_default_formulas(
        start_date=start_date,
        end_date=end_date,
        fiscal_year=fiscal_year,
        quarter_dates=quarter_dates,
        previous_quarter_dates=previous_quarter_dates,
        quarter_criteria=quarter_criteria,
        previous_quarter_criteria=previous_quarter_criteria,
        start_exclusive=start_exclusive,
        end_inclusive=end_inclusive,
    )


def _investment_case_scenario_render_deps(wb: Workbook) -> InvestmentCaseScenarioRenderDeps:
    return InvestmentCaseScenarioRenderDeps(runtime={"wb": wb})


def _write_scenario_driver_assumptions_sheet(
    wb: Workbook,
    *,
    ticker: Any,
    segment_specs: Sequence[_SegmentScenarioInputSpec] = (),
    enabled: bool = True,
    disabled_note: str = "",
) -> None:
    return _write_scenario_driver_assumptions_sheet_impl(
        _investment_case_scenario_render_deps(wb),
        ticker=ticker,
        segment_specs=segment_specs,
        enabled=enabled,
        disabled_note=disabled_note,
    )


def _write_scenario_bridge_tax_treatment_sheet(
    wb: Workbook,
    *,
    ticker: Any,
    specs: Sequence[_ScenarioDriverBridgeSpec],
    after_tax_factor: Optional[float] = None,
    tax_rate_ref: str = "",
    tax_source_basis: str = "",
) -> None:
    return _write_scenario_bridge_tax_treatment_sheet_impl(
        _investment_case_scenario_render_deps(wb),
        ticker=ticker,
        specs=specs,
        after_tax_factor=after_tax_factor,
        tax_rate_ref=tax_rate_ref,
        tax_source_basis=tax_source_basis,
    )



def _operating_driver_workbook_support_runtime() -> Dict[str, Any]:
    return {
        "re": re,
        "date": date,
        "_date_or_none": _date_or_none,
    }


def _operating_driver_workbook_support() -> OperatingDriverWorkbookSupport:
    return OperatingDriverWorkbookSupport(
        OperatingDriverWorkbookSupportDeps(runtime=_operating_driver_workbook_support_runtime())
    )


def _operating_driver_ttm_sum_from_workbook(wb: Workbook, metric_label: str) -> Optional[float]:
    return _operating_driver_workbook_support().operating_driver_ttm_sum_from_workbook(wb, metric_label)


def _operating_driver_latest_full_year_sum_from_workbook(wb: Workbook, metric_label: str) -> Optional[float]:
    return _operating_driver_workbook_support().operating_driver_latest_full_year_sum_from_workbook(
        wb,
        metric_label,
    )




def _anf_visible_support_runtime() -> Dict[str, Any]:
    return {
        "pd": pd,
        "re": re,
        "math": math,
        "date": date,
        "timedelta": timedelta,
        "glx_normalize_text": glx_normalize_text,
        "_shared_visible_period_text": _shared_visible_period_text,
        "_promise_metric_definition_key": _promise_metric_definition_key,
    }


def _anf_visible_support() -> AnfVisibleSupport:
    return AnfVisibleSupport(AnfVisibleSupportDeps(runtime=_anf_visible_support_runtime()))


def _anf_fiscal_year_from_quarter_end(qd: Any) -> Optional[int]:
    return _anf_visible_support().fiscal_year_from_quarter_end(qd)


def _anf_fiscal_quarter_from_quarter_end(qd: Any) -> Optional[int]:
    return _anf_visible_support().fiscal_quarter_from_quarter_end(qd)


def _anf_visible_quarter_label(qd: Any) -> str:
    return _anf_visible_support().visible_quarter_label(qd)


def _source_backed_debt_tranches_from_slides(
    slides_debt: Optional[pd.DataFrame],
    latest_quarter: Any,
) -> pd.DataFrame:
    return source_backed_debt_tranches_from_slides(
        ValuationDebtSupportDeps(runtime={"pd": pd, "re": re}),
        slides_debt,
        latest_quarter,
    )


def _shared_visible_period_text(text_in: Any) -> str:
    """Normalize user-facing period labels without touching source dates."""
    txt = str(text_in or "")
    if not txt:
        return ""
    txt = re.sub(r"\bQ([1-4])\s+FY\s*(20\d{2})\b", r"\2-Q\1", txt, flags=re.I)
    txt = re.sub(r"\bQ([1-4])\s+fiscal\s+(20\d{2})\b", r"\2-Q\1", txt, flags=re.I)
    txt = re.sub(r"\bQ([1-4])\s*[-/]\s*(20\d{2})\b", r"\2-Q\1", txt, flags=re.I)
    txt = re.sub(r"\bQ([1-4])\s+(20\d{2})\b", r"\2-Q\1", txt, flags=re.I)
    txt = re.sub(r"\bFY\s*(20\d{2})\b", r"\1 year", txt, flags=re.I)
    txt = re.sub(r"\bfiscal\s+year\s+(20\d{2})\b", r"\1 year", txt, flags=re.I)
    return txt


def _shared_readable_source_label(text_in: Any) -> str:
    txt = str(text_in or "")
    if not txt:
        return ""
    replacements = [
        (r"\brel\s*/\s*qtr\s*hist\.?\b", "earnings release / quarterly history"),
        (r"\brel\s*/\s*qtr\b", "earnings release / quarterly history"),
        (r"\bqtr\s*hist\.?\b", "quarterly history"),
        (r"\brel\s*/\s*tr\.?\b", "earnings release / earnings transcript"),
        (r"\btr\.?\b", "earnings transcript"),
        (r"\bsec\b", "SEC filing"),
        (r"\bslides\b", "investor presentation"),
    ]
    out = txt
    for pattern, repl in replacements:
        out = re.sub(pattern, repl, out, flags=re.I)
    return out


def _shared_readable_source_type_label(text_in: Any) -> str:
    """Display compact internal source ids as reviewer-readable labels."""
    txt = str(text_in or "").strip()
    if not txt:
        return ""
    key = re.sub(r"[^a-z0-9]+", "_", txt.lower()).strip("_")
    mapping = {
        "adj_metrics": "adjusted metrics",
        "earnings_presentation": "earnings presentation",
        "earnings_release": "earnings release",
        "financial_statement": "financial schedule",
        "history_q": "quarterly history",
        "model_metric": "model-derived",
        "modeled": "model-derived",
        "promise": "management guidance",
        "promise_text": "management guidance",
        "quarter_note": "quarter notes",
        "quarter_notes": "quarter notes",
        "quarter_notes_ui": "quarter notes",
        "slides_debt_profile": "debt profile slides",
        "transcript": "earnings transcript",
    }
    if key in mapping:
        return mapping[key]
    return re.sub(r"[_-]+", " ", txt).strip()


def _sector_operating_driver_intro_tables(ticker: Any) -> List[Dict[str, Any]]:
    return SectorOperatingDriverIntroSupport(
        SectorOperatingDriverIntroSupportDeps(
            runtime={
                "_guidance_source_contract_label": _guidance_source_contract_label,
            }
        )
    ).sector_operating_driver_intro_tables(ticker)


def _standardize_quarter_notes_ui_categories(ws: Any, ticker: Any) -> None:
    return _standardize_quarter_notes_ui_categories_impl(
        QuarterNotesContextAdapterDeps(
            runtime={
                "_shared_visible_period_text": _shared_visible_period_text,
                "glx_normalize_text": glx_normalize_text,
            }
        ),
        ws,
        ticker=ticker,
    )


def _rewrite_shared_promise_progress_ui_from_blocks(ws: Any, ticker: Any = "") -> None:
    return rewrite_shared_promise_progress_ui_from_blocks(
        PromiseProgressRewriteDeps(runtime={**globals(), **locals()}),
        ws,
        ticker=ticker,
    )




def _promise_progress_worksheet_repair_deps() -> PromiseProgressWorksheetRepairDeps:
    return PromiseProgressWorksheetRepairDeps(runtime={**globals()})


def _management_credibility_scorecard_rows(ticker: Any = "") -> List[Tuple[str, str, str, str]]:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._management_credibility_scorecard_rows(ticker)


def _insert_management_credibility_scorecard(ws: Any, ticker: Any = "") -> None:
    return insert_management_credibility_scorecard(
        _promise_progress_worksheet_repair_deps(),
        ws,
        ticker,
    )


def _polish_promise_scorecard_layout(ws: Any) -> None:
    return polish_promise_scorecard_layout(
        _promise_progress_worksheet_repair_deps(),
        ws,
    )


def _promise_header_name(value: Any) -> str:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._promise_header_name(value)


def _set_promise_row_semantics(
    ws: Any,
    row_idx: int,
    cols: Mapping[str, int],
    *,
    change_type: Any = None,
    actual: Any = None,
    progress: Any = None,
    status: Any = None,
    note: Any = None,
) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._set_promise_row_semantics(
        ws,
        row_idx,
        cols,
        change_type=change_type,
        actual=actual,
        progress=progress,
        status=status,
        note=note,
    )


def _promise_hidden_key_slug(value: Any) -> str:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._promise_hidden_key_slug(value)


def _ensure_anf_promise_hidden_source_keys(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._ensure_anf_promise_hidden_source_keys(ws)


def _promise_stated_quarter_parts(value: Any) -> Tuple[Optional[int], Optional[int]]:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._promise_stated_quarter_parts(value)


def _promise_annual_year(value: Any) -> Optional[int]:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._promise_annual_year(value)


def _promise_progress_label(value: Any, *, metric: Any = "", stated: Any = "") -> str:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._promise_progress_label(value, metric=metric, stated=stated)


def _promise_value_looks_like_progress(value: Any, metric: Any = "") -> bool:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._promise_value_looks_like_progress(value, metric=metric)


def _promise_revision_event_from_section(section: Any) -> str:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._promise_revision_event_from_section(section)


def _promise_event_sort_key(value: Any, source_date: Any = "") -> Tuple[date, int, str]:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._promise_event_sort_key(value, source_date)


def _promise_metric_definition_key(metric: Any) -> str:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._promise_metric_definition_key(metric)


def _promise_metric_order_rank(metric: Any) -> int:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._promise_metric_order_rank(metric)


def _clean_gpre_45z_monetization_value(value: Any) -> Any:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._clean_gpre_45z_monetization_value(value)


def _finalize_promise_revision_semantics(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._finalize_promise_revision_semantics(ws)


def _promise_status_fill_for_label(status: Any) -> PatternFill:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._promise_status_fill_for_label(status)


def _apply_promise_grid_style(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._apply_promise_grid_style(ws)


def _cleanup_anf_promise_after_repair(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._cleanup_anf_promise_after_repair(ws)


def _repair_anf_promise_actual_progress_semantics(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._repair_anf_promise_actual_progress_semantics(ws)


def _clear_pre_release_promise_actuals(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._clear_pre_release_promise_actuals(ws)


def _remove_empty_promise_revision_blocks(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._remove_empty_promise_revision_blocks(ws)


def _standardize_promise_section_layout(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._standardize_promise_section_layout(ws)


def _is_promise_section_row(ws: Any, row_idx: int) -> bool:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._is_promise_section_row(ws, row_idx)


def _ensure_q4_annual_actual_revision_rows(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._ensure_q4_annual_actual_revision_rows(ws)


def _ensure_promise_block_spacing(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._ensure_promise_block_spacing(ws)


def _dedupe_promise_progress_rows(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._dedupe_promise_progress_rows(ws)


def _remove_actual_only_promise_rows(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._remove_actual_only_promise_rows(ws)


def _remove_promise_metric_stubs(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._remove_promise_metric_stubs(ws)


def _remove_pbi_duplicate_cost_savings_timeline_rows(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._remove_pbi_duplicate_cost_savings_timeline_rows(ws)


def _ensure_cost_savings_run_rate_revision_row(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._ensure_cost_savings_run_rate_revision_row(ws)


def _remove_blank_promise_rows(ws: Any) -> None:
    _promise_progress_worksheet_repairs._set_runtime(_promise_progress_worksheet_repair_deps())
    return _promise_progress_worksheet_repairs._remove_blank_promise_rows(ws)


def _repair_promise_table_header_merges(ws: Any) -> None:
    return repair_promise_table_header_merges(
        _promise_progress_worksheet_repair_deps(),
        ws,
    )


def _final_repair_promise_progress_ui(wb: Workbook, ticker: Any = "") -> None:
    return final_repair_promise_progress_ui(
        _promise_progress_worksheet_repair_deps(),
        wb,
        ticker,
    )






































def _polish_investment_case_readability(ws: Any) -> None:
    return InvestmentCaseReadability(
        InvestmentCaseReadabilityDeps(runtime={"Alignment": Alignment})
    ).polish_investment_case_readability(ws)




































def _apply_source_backed_promise_mapping_overrides(wb: Workbook, ticker: Any = "") -> None:
    return apply_source_backed_promise_mapping_overrides(
        PromiseSourceOverrideMutatorDeps(runtime={**globals(), **locals()})
    )


def _shared_ui_conventions_deps() -> SharedUiConventionsDeps:
    return SharedUiConventionsDeps(
        runtime={
            "PatternFill": PatternFill,
            "Border": Border,
            "Side": Side,
            "Font": Font,
            "Alignment": Alignment,
            "Comment": Comment,
            "get_column_letter": get_column_letter,
            "copy": copy,
            "pd": pd,
            "math": math,
            "re": re,
            "date": date,
            "_shared_visible_period_text": _shared_visible_period_text,
            "_shared_readable_source_label": _shared_readable_source_label,
            "_standardize_quarter_notes_ui_categories": _standardize_quarter_notes_ui_categories,
            "_remove_empty_promise_revision_blocks": _remove_empty_promise_revision_blocks,
            "_polish_promise_scorecard_layout": _polish_promise_scorecard_layout,
            "_apply_source_backed_promise_mapping_overrides": _apply_source_backed_promise_mapping_overrides,
            "_polish_investment_case_readability": _polish_investment_case_readability,
            "_date_or_none": _date_or_none,
            "_promise_progress_label": _promise_progress_label,
            "PROMISE_TIMELINE_HEADERS": PROMISE_TIMELINE_HEADERS,
            "PROMISE_VISIBLE_MAX_COL": PROMISE_VISIBLE_MAX_COL,
        }
    )


def _apply_shared_ui_conventions_to_workbook(wb: Workbook, ticker: Any = "") -> None:
    return apply_shared_ui_conventions_to_workbook(
        _shared_ui_conventions_deps(),
        wb,
        ticker=ticker,
    )


def _anf_valuation_support_runtime() -> Dict[str, Any]:
    return {
        "pd": pd,
        "np": np,
        "dt": dt,
        "_anf_fiscal_year_from_quarter_end": _anf_fiscal_year_from_quarter_end,
        "_anf_fiscal_quarter_from_quarter_end": _anf_fiscal_quarter_from_quarter_end,
        "_anf_visible_quarter_label": _anf_visible_quarter_label,
    }


def _anf_valuation_support() -> AnfValuationSupport:
    return AnfValuationSupport(AnfValuationSupportDeps(runtime=_anf_valuation_support_runtime()))


def _anf_buyback_execution_is_year_or_ttm(
    qd: Any,
    note_text: Any = "",
    *,
    cash_amount: Optional[float] = None,
    shares_amount: Optional[float] = None,
) -> bool:
    return _anf_valuation_support().buyback_execution_is_year_or_ttm(
        qd,
        note_text,
        cash_amount=cash_amount,
        shares_amount=shares_amount,
    )


def _anf_format_year_ttm_buyback_summary(
    qd: Any,
    *,
    shares_amount: Optional[float] = None,
    cash_amount: Optional[float] = None,
    avg_price: Optional[float] = None,
) -> str:
    return _anf_valuation_support().format_year_ttm_buyback_summary(
        qd,
        shares_amount=shares_amount,
        cash_amount=cash_amount,
        avg_price=avg_price,
    )


def _anf_normalized_quarter_ts(qd: Any) -> Optional[pd.Timestamp]:
    return _anf_valuation_support().normalized_quarter_ts(qd)


def _anf_quarter_sequence(quarters: Iterable[Any]) -> List[pd.Timestamp]:
    return _anf_valuation_support().quarter_sequence(quarters)


def _anf_prior_year_quarter(qd: Any, quarters: Iterable[Any]) -> Optional[pd.Timestamp]:
    return _anf_valuation_support().prior_year_quarter(qd, quarters)


def _anf_previous_quarter(qd: Any, quarters: Iterable[Any]) -> Optional[pd.Timestamp]:
    return _anf_valuation_support().previous_quarter(qd, quarters)


def _anf_normalize_value_map(src: Dict[Any, Any]) -> Dict[pd.Timestamp, Any]:
    return _anf_valuation_support().normalize_value_map(src)


def _anf_is_missing_value(v: Any) -> bool:
    return _anf_valuation_support().is_missing_value(v)


def _anf_yoy_map_for_fiscal_periods(
    src: Dict[Any, Any],
    quarters: Iterable[Any],
    *,
    positive_prev_only: bool = False,
    positive_cur_only: bool = False,
) -> Dict[pd.Timestamp, Any]:
    return _anf_valuation_support().yoy_map_for_fiscal_periods(
        src,
        quarters,
        positive_prev_only=positive_prev_only,
        positive_cur_only=positive_cur_only,
    )


def _anf_value_delta_map_for_fiscal_periods(
    src: Dict[Any, Any],
    quarters: Iterable[Any],
    *,
    comparison: str = "yoy",
) -> Dict[pd.Timestamp, Any]:
    return _anf_valuation_support().value_delta_map_for_fiscal_periods(
        src,
        quarters,
        comparison=comparison,
    )


def _anf_normalize_ytd_buyback_cash_map_for_valuation(
    src: Dict[Any, Any],
    quarters: Iterable[Any],
) -> Dict[pd.Timestamp, Any]:
    return _anf_valuation_support().normalize_ytd_buyback_cash_map_for_valuation(src, quarters)


def _anf_format_guidance_display_value(metric: Any, low: Any, high: Any, value: Any, unit: Any, line: Any = "") -> str:
    return _anf_visible_support().format_guidance_display_value(metric, low, high, value, unit, line)


def _anf_valuation_guidance_rows(guidance_df: pd.DataFrame) -> List[Dict[str, str]]:
    return _anf_visible_support().valuation_guidance_rows(guidance_df)


ANF_SEGMENT_BRAND_EXPLANATION = (
    "Americas / EMEA / APAC are geographic segments; Abercrombie / Hollister are brand families. "
    "Total Company is not additive across both views."
)


def _anf_valuation_side_panel_runtime() -> Dict[str, Any]:
    return {
        "copy": copy,
        "get_column_letter": get_column_letter,
        "PatternFill": PatternFill,
        "Border": Border,
        "Side": Side,
        "Font": Font,
        "Alignment": Alignment,
        "Comment": Comment,
        "_anf_clean_visible_ui_text": _anf_clean_visible_ui_text,
        "_shared_visible_period_text": _shared_visible_period_text,
        "_style_valuation_side_panel_style_bundle": _style_valuation_side_panel_style_bundle,
    }


def _anf_valuation_side_panel_deps() -> AnfValuationSidePanelDeps:
    return AnfValuationSidePanelDeps(runtime=_anf_valuation_side_panel_runtime())


def _anf_clear_valuation_side_panels(
    ws: Any,
    *,
    start_col: int = 15,
    end_col: Optional[int] = None,
    side_max_row: int = 125,
) -> None:
    return clear_anf_valuation_side_panels(
        _anf_valuation_side_panel_deps(),
        ws,
        start_col=start_col,
        end_col=end_col,
        side_max_row=side_max_row,
    )


def _valuation_side_panel_style_bundle() -> Dict[str, Any]:
    return valuation_side_panel_style_bundle(_anf_valuation_side_panel_deps())


def _write_anf_valuation_side_panel(ws: Any, *, start_row: int = 7, start_col: int = 15, end_col: int = 29) -> Dict[str, int]:
    return write_anf_valuation_side_panel(
        _anf_valuation_side_panel_deps(),
        ws,
        start_row=start_row,
        start_col=start_col,
        end_col=end_col,
    )


def _anf_clean_visible_ui_text(text_in: Any, *, max_chars: Optional[int] = None) -> str:
    return _anf_visible_support().clean_visible_ui_text(text_in, max_chars=max_chars)


def _anf_visible_quarter_note_summaries(
    text_in: Any,
    *,
    quarter_label: Any = "",
    latest_label: Any = "",
) -> List[str]:
    return _anf_visible_support().visible_quarter_note_summaries(text_in, quarter_label=quarter_label, latest_label=latest_label)


def _anf_compact_driver_label(label_in: Any, unit_txt: Any = "") -> str:
    return _anf_visible_support().compact_driver_label(label_in, unit_txt)


def _anf_compact_driver_group(group_in: Any, label_in: Any = "", driver_key: Any = "") -> str:
    return _anf_visible_support().compact_driver_group(group_in, label_in, driver_key)


def _anf_round_visible_driver_value(value_in: Any, unit_txt: Any = "", label_in: Any = "", driver_key: Any = "") -> Optional[float]:
    return _anf_visible_support().round_visible_driver_value(value_in, unit_txt, label_in, driver_key)


def _anf_polish_quarter_note_visible_fields(category_in: Any, metric_in: Any, note_in: Any) -> Tuple[str, str]:
    return _anf_visible_support().polish_quarter_note_visible_fields(category_in, metric_in, note_in)


def _anf_clean_visible_operating_driver_records(rows_in: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return _anf_visible_support().clean_visible_operating_driver_records(rows_in)


def _investment_case_support_runtime() -> Dict[str, Any]:
    return {
        "pd": pd,
        "math": math,
        "re": re,
        "date": date,
        "glx_normalize_text": glx_normalize_text,
        "_anf_visible_guidance_normalized_frame": _anf_visible_guidance_normalized_frame,
        "_anf_visible_quarter_label": _anf_visible_quarter_label,
        "_guidance_source_contract_label": _guidance_source_contract_label,
        "_segment_scenario_revenue_m": _segment_scenario_revenue_m,
        "_shared_visible_period_text": _shared_visible_period_text,
        "_anf_clean_visible_ui_text": _anf_clean_visible_ui_text,
    }


def _investment_case_support() -> InvestmentCaseSupport:
    return InvestmentCaseSupport(
        InvestmentCaseSupportDeps(runtime=_investment_case_support_runtime())
    )


def _anf_investment_case_sheet_order(
    desired_sheet_order: Sequence[str],
    raw_sheet_cluster: Sequence[str],
    *,
    is_anf_profile: bool = False,
) -> Tuple[Tuple[str, ...], Tuple[str, ...]]:
    return _investment_case_support().anf_investment_case_sheet_order(
        desired_sheet_order,
        raw_sheet_cluster,
        is_anf_profile=is_anf_profile,
    )


def _investment_case_sheet_order(
    desired_sheet_order: Sequence[str],
    raw_sheet_cluster: Sequence[str],
    *,
    ticker: Any = "",
) -> Tuple[Tuple[str, ...], Tuple[str, ...]]:
    return _investment_case_support().investment_case_sheet_order(
        desired_sheet_order,
        raw_sheet_cluster,
        ticker=ticker,
    )


def _anf_build_investment_case_data(
    *,
    hist: Any,
    operating_driver_rows: Sequence[Dict[str, Any]],
    guidance_normalized: Any,
    slides_segments: Any,
    valuation_summary: Any = None,
    adjusted_metrics: Any = None,
) -> pd.DataFrame:
    return _investment_case_support().build_anf_investment_case_data(
        hist=hist,
        operating_driver_rows=operating_driver_rows,
        guidance_normalized=guidance_normalized,
        slides_segments=slides_segments,
        valuation_summary=valuation_summary,
        adjusted_metrics=adjusted_metrics,
    )


def _sector_build_investment_case_data(
    *,
    ticker: str,
    hist: Any,
    operating_driver_rows: Sequence[Dict[str, Any]] = (),
    guidance_normalized: Any = None,
    valuation_summary: Any = None,
    economics_market_rows: Any = None,
    slides_segments: Any = None,
) -> pd.DataFrame:
    return _investment_case_support().build_sector_investment_case_data(
        ticker=ticker,
        hist=hist,
        operating_driver_rows=operating_driver_rows,
        guidance_normalized=guidance_normalized,
        valuation_summary=valuation_summary,
        economics_market_rows=economics_market_rows,
        slides_segments=slides_segments,
    )



def _sector_investment_case_render_deps(wb: Workbook) -> SectorInvestmentCaseRenderDeps:
    return SectorInvestmentCaseRenderDeps(
        runtime={
            "wb": wb,
            "SCENARIO_DRIVER_CAPITAL_STRUCTURE_INTEREST": SCENARIO_DRIVER_CAPITAL_STRUCTURE_INTEREST,
            "SCENARIO_DRIVER_CASH_FLOW_CAPEX": SCENARIO_DRIVER_CASH_FLOW_CAPEX,
            "SCENARIO_DRIVER_MANUAL_INCREMENTAL": SCENARIO_DRIVER_MANUAL_INCREMENTAL,
            "SCENARIO_DRIVER_MARGIN_EBITDA": SCENARIO_DRIVER_MARGIN_EBITDA,
            "SCENARIO_DRIVER_TAX_CREDIT_SUBSIDY": SCENARIO_DRIVER_TAX_CREDIT_SUBSIDY,
            "SCENARIO_TAX_CASH_ONLY": SCENARIO_TAX_CASH_ONLY,
            "SCENARIO_TAX_NON_TAXABLE_CREDIT": SCENARIO_TAX_NON_TAXABLE_CREDIT,
            "SCENARIO_TAX_NO_EPS_IMPACT": SCENARIO_TAX_NO_EPS_IMPACT,
            "SCENARIO_TAX_TAXABLE": SCENARIO_TAX_TAXABLE,
            "_ScenarioDriverBridgeSpec": _ScenarioDriverBridgeSpec,
            "_SegmentScenarioInputSpec": _SegmentScenarioInputSpec,
            "_bs_segments_latest_segment_margin_from_workbook": _bs_segments_latest_segment_margin_from_workbook,
            "_company_operating_margin_proxy_from_workbook": _company_operating_margin_proxy_from_workbook,
            "_date_or_none": _date_or_none,
            "_excel_manual_percent_active_formula": _excel_manual_percent_active_formula,
            "_excel_percent_value_expr": _excel_percent_value_expr,
            "_excel_visible_value_range_formula": _excel_visible_value_range_formula,
            "_fiscal_profile_from_workbook": _fiscal_profile_from_workbook,
            "_guidance_source_contract_label": _guidance_source_contract_label,
            "_history_q_latest_full_year_actuals_from_workbook": _history_q_latest_full_year_actuals_from_workbook,
            "_history_q_latest_full_year_period_set": _history_q_latest_full_year_period_set,
            "_history_q_year_default_formulas": _history_q_year_default_formulas,
            "_operating_driver_latest_full_year_sum_from_workbook": _operating_driver_latest_full_year_sum_from_workbook,
            "_operating_driver_ttm_sum_from_workbook": _operating_driver_ttm_sum_from_workbook,
            "_scenario_bridge_eps_value_formula": _scenario_bridge_eps_value_formula,
            "_scenario_bridge_row_values": _scenario_bridge_row_values,
            "_segment_scenario_specs_from_records": _segment_scenario_specs_from_records,
            "_shared_visible_period_text": _shared_visible_period_text,
            "_write_scenario_bridge_tax_treatment_sheet": _write_scenario_bridge_tax_treatment_sheet,
            "_write_scenario_driver_assumptions_sheet": _write_scenario_driver_assumptions_sheet,
        }
    )


def _write_sector_investment_case_data_sheet(wb: Workbook, ticker: Any, data: pd.DataFrame) -> None:
    return write_sector_investment_case_data_sheet(
        _sector_investment_case_render_deps(wb),
        data,
        ticker=str(ticker or "").strip().upper(),
    )


def _write_sector_investment_case_sheet(wb: Workbook, ticker: Any, data: pd.DataFrame) -> None:
    return write_sector_investment_case_sheet(
        _sector_investment_case_render_deps(wb),
        data,
        ticker=str(ticker or "").strip().upper(),
    )


def _anf_investment_case_render_deps(wb: Workbook) -> AnfInvestmentCaseRenderDeps:
    return AnfInvestmentCaseRenderDeps(
        runtime={
            "wb": wb,
            "SCENARIO_DRIVER_CASH_FLOW_CAPEX": SCENARIO_DRIVER_CASH_FLOW_CAPEX,
            "SCENARIO_DRIVER_MARGIN_EBITDA": SCENARIO_DRIVER_MARGIN_EBITDA,
            "SCENARIO_DRIVER_SHARE_COUNT_BUYBACK": SCENARIO_DRIVER_SHARE_COUNT_BUYBACK,
            "SCENARIO_TAX_CASH_ONLY": SCENARIO_TAX_CASH_ONLY,
            "SCENARIO_TAX_NO_EPS_IMPACT": SCENARIO_TAX_NO_EPS_IMPACT,
            "SCENARIO_TAX_TAXABLE": SCENARIO_TAX_TAXABLE,
            "_ScenarioDriverBridgeSpec": _ScenarioDriverBridgeSpec,
            "_SegmentScenarioInputSpec": _SegmentScenarioInputSpec,
            "_anf_clean_visible_ui_text": _anf_clean_visible_ui_text,
            "_excel_manual_percent_active_formula": _excel_manual_percent_active_formula,
            "_excel_percent_value_expr": _excel_percent_value_expr,
            "_excel_visible_value_range_formula": _excel_visible_value_range_formula,
            "_history_q_latest_full_year_actuals_from_workbook": _history_q_latest_full_year_actuals_from_workbook,
            "_history_q_latest_full_year_period_set": _history_q_latest_full_year_period_set,
            "_history_q_year_default_formulas": _history_q_year_default_formulas,
            "_scenario_bridge_eps_value_formula": _scenario_bridge_eps_value_formula,
            "_scenario_bridge_row_values": _scenario_bridge_row_values,
            "_segment_scenario_revenue_m": _segment_scenario_revenue_m,
            "_segment_scenario_specs_from_records": _segment_scenario_specs_from_records,
            "_segment_scenario_view_basis": _segment_scenario_view_basis,
            "_write_scenario_bridge_tax_treatment_sheet": _write_scenario_bridge_tax_treatment_sheet,
            "_write_scenario_driver_assumptions_sheet": _write_scenario_driver_assumptions_sheet,
        }
    )


def _write_anf_investment_case_data_sheet(wb: Workbook, data: pd.DataFrame) -> None:
    return write_anf_investment_case_data_sheet(
        _anf_investment_case_render_deps(wb),
        data,
    )


def _write_anf_investment_case_sheet(wb: Workbook, data: pd.DataFrame) -> None:
    return write_anf_investment_case_sheet(
        _anf_investment_case_render_deps(wb),
        data,
    )
















def _anf_guidance_visible_period_label(period_label: Any, source_quarter: Any = None) -> str:
    return _anf_visible_support().guidance_visible_period_label(period_label, source_quarter)


def _anf_guidance_metric_unit_is_compatible(metric_hint: Any, unit: Any, line: Any = "") -> bool:
    return _anf_visible_support().guidance_metric_unit_is_compatible(metric_hint, unit, line)


def _anf_guidance_horizon_type(label_in: Any) -> str:
    return _anf_visible_support().guidance_horizon_type(label_in)


def _anf_reclassify_guidance_period_label(period_label: Any, metric_hint: Any, line: Any) -> str:
    return _anf_visible_support().reclassify_guidance_period_label(period_label, metric_hint, line)


def _anf_visible_guidance_normalized_frame(guidance_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    return _anf_visible_support().visible_guidance_normalized_frame(guidance_df)


def _anf_format_guidance_value(metric: str, low: Any = None, high: Any = None, value: Any = None, unit: Any = "", numbers: Any = "") -> str:
    return _anf_visible_support().format_guidance_value(metric, low, high, value, unit, numbers)


def _anf_build_guidance_timeline_rows(guidance_df: Optional[pd.DataFrame] = None, hist_df: Optional[pd.DataFrame] = None) -> List[Dict[str, str]]:
    return _anf_visible_support().build_guidance_timeline_rows(guidance_df, hist_df)


def _anf_build_promise_progress_sections(guidance_df: Optional[pd.DataFrame], hist_df: Optional[pd.DataFrame] = None) -> Dict[str, List[Dict[str, str]]]:
    return _anf_visible_support().build_promise_progress_sections(guidance_df, hist_df)


def _anf_recent_operating_commentary_rows(
    hist_df: Optional[pd.DataFrame],
    slides_segments: Optional[pd.DataFrame],
    quarters: Sequence[Any],
) -> List[Dict[str, Any]]:
    return _anf_visible_support().recent_operating_commentary_rows(hist_df, slides_segments, quarters)




def _slides_guidance_metric_key(metric_name: Any) -> str:
    return _anf_visible_support().slides_guidance_metric_key(metric_name)


def _slides_guidance_has_explicit_metric(
    slides_guidance: pd.DataFrame,
    qd: date,
    metric_name: str,
    *,
    require_range: bool = False,
) -> bool:
    return _anf_visible_support().slides_guidance_has_explicit_metric(slides_guidance, qd, metric_name, require_range=require_range)


def _anf_financial_schedule_support_doc_for_quarter(
    qd: date,
    *,
    adj_metrics: pd.DataFrame,
    non_gaap_files: pd.DataFrame,
    slides_segments: pd.DataFrame,
) -> str:
    return _anf_visible_support().financial_schedule_support_doc_for_quarter(qd, adj_metrics=adj_metrics, non_gaap_files=non_gaap_files, slides_segments=slides_segments)


try:
    from bs4 import BeautifulSoup
except Exception:  # pragma: no cover - optional dependency in this pipeline
    BeautifulSoup = None

from .cache_layout import canonical_shared_cache_root, ticker_cache_candidates, ticker_cache_roots_from_base_dir
from .path_config import data_root_from_sec_cache_path
from .capital_return_notes import (
    build_buyback_note as capital_return_build_buyback_note,
    build_dividend_note as capital_return_build_dividend_note,
    build_dividend_note_from_text as capital_return_build_dividend_note_from_text,
    normalize_capital_return_note_item as capital_return_normalize_note_item,
    normalize_new_prefix as capital_return_normalize_new_prefix,
)
from .company_profiles import COMPANY_PROFILES, get_company_profile
from .debt_parser import coerce_number, read_html_tables_any
from .filing_evidence_shared import (
    classify_statement_evidence_role as shared_classify_statement_evidence_role,
    build_canonical_subject_key as shared_build_canonical_subject_key,
    build_evidence_event as shared_build_evidence_event,
    build_follow_through_event as shared_build_follow_through_event,
    build_follow_through_signal as shared_build_follow_through_signal,
    investor_note_candidate_from_event as shared_investor_note_candidate_from_event,
    build_lifecycle_subject_key as shared_build_lifecycle_subject_key,
    build_parent_subject_key as shared_build_parent_subject_key,
    build_promise_lifecycle_key as shared_build_promise_lifecycle_key,
    derive_lifecycle_state as shared_derive_lifecycle_state,
    derive_status_resolution_reason as shared_derive_status_resolution_reason,
    evidence_role as shared_evidence_role,
    infer_target_period_norm as shared_infer_target_period_norm,
    is_preferred_narrative_source as shared_is_preferred_narrative_source,
    merge_same_subject_events as shared_merge_same_subject_events,
    merge_follow_through_signals as shared_merge_follow_through_signals,
    merge_evidence_events as shared_merge_evidence_events,
    narrative_drop_reason as shared_narrative_drop_reason,
    pick_best_subject_row_for_quarter as shared_pick_best_subject_row_for_quarter,
    progress_status_rank as shared_progress_status_rank,
    promise_candidate_drop_reason as shared_promise_candidate_drop_reason,
    qualify_promise_candidate as shared_qualify_promise_candidate,
    qualify_renderable_note as shared_qualify_renderable_note,
    renderable_note_drop_reason as shared_renderable_note_drop_reason,
    route_to_measurable_promise_candidate as shared_route_to_measurable_promise_candidate,
    source_class as shared_source_class,
    statement_class as shared_statement_class,
    looks_like_tabular_fragment as shared_looks_like_tabular_fragment,
)
from .doc_intel import extract_pdf_text_cached
from .derivative_crush_tests import build_derivative_crush_tests
from .derivative_oci_bridge import DERIVATIVE_EXPOSURE_COLUMNS, build_derivative_oci_bridge_from_sources
from .excel_writer_drivers import (
    candidate_records_for_template as driver_candidate_records_for_template,
    driver_best_text_record as driver_driver_best_text_record,
    driver_snippet as driver_driver_snippet,
    driver_source_display as driver_source_display,
    group_operating_driver_source_records_by_quarter,
    load_operating_driver_45z_guidance_docs_by_quarter as driver_load_operating_driver_45z_guidance_docs_by_quarter,
    load_operating_driver_bridge_bundle_map as driver_load_operating_driver_bridge_bundle_map,
    load_operating_driver_source_records as driver_load_operating_driver_source_records,
    load_operating_driver_template_index as driver_load_operating_driver_template_index,
    operating_driver_order_map as driver_operating_driver_order_map,
    operating_driver_template_spec as driver_operating_driver_template_spec,
    build_operating_driver_line_index as driver_build_operating_driver_line_index,
    template_candidate_terms as driver_template_candidate_terms,
    text_matches_template_terms as driver_text_matches_template_terms,
)
from .excel_writer_basis_proxy_sandbox import (
    BasisProxySandboxWriterDeps,
    write_basis_proxy_sandbox_sheet,
)
from .excel_writer_economics_overlay import (
    GpreOverlayQuarterComparisonDeps,
    GpreOverlaySupportInputs,
    _overlay_model_label,
    write_gpre_overlay_quarter_comparisons,
    write_gpre_basis_proxy_overlay_support,
)
from .excel_writer_economics_overlay_derivatives import (
    GpreEconomicsOverlayDerivativeSideEffectDeps,
    write_gpre_derivative_crush_tests_side_effect,
)
from .excel_writer_derivative_oci_bridge import (
    DerivativeOciBridgeRenderDeps,
    write_derivative_crush_tests_sheet,
    write_derivative_oci_bridge_sheet,
)
from .excel_writer_debt_convertible_enrichment import (
    DebtConvertibleEnrichmentDeps,
    DebtConvertibleEnrichmentSupport,
)
from .excel_writer_summary_sheet import (
    SummarySheetRenderDeps,
    write_summary_sheet,
)
from .excel_writer_hidden_value_flags import (
    HiddenValueFlagsSheetInputs,
    write_hidden_value_flags_sheet,
)
from .excel_writer_hidden_value_support import (
    HiddenValueSupport,
    HiddenValueSupportDeps,
)
from .excel_writer_hidden_value_surface import (
    HiddenValueSurfaceModelInputs,
    NO_TRIGGER_DISPLAY_LABEL,
    NO_TRIGGER_DISPLAY_SCORE,
    NO_TRIGGER_DISPLAY_SEVERITY,
    NO_TRIGGER_DISPLAY_SUPPORT,
    NO_TRIGGER_DISPLAY_TITLE,
    build_hidden_value_surface_model,
    hidden_flag_field,
    hidden_flag_score,
    hidden_value_ai_helper_formula,
)
from .excel_writer_promise_progress import (
    PromiseProgressRenderHelpers,
    PromiseProgressSheetInputs,
    write_promise_progress_sheet,
)
from .excel_writer_promise_progress_anf import (
    AnfPromiseProgressWriterDeps,
    write_anf_promise_progress_ui_sheet,
)
from .excel_writer_promise_progress_render_adapter import (
    PromiseProgressRowWriterDeps,
    build_promise_progress_row_writer,
)
from .excel_writer_promise_progress_sources import (
    PromiseProgressSourceDeps,
    PromiseProgressSourceSupport,
)
from .excel_writer_promise_progress_bundle import (
    PromiseProgressUiBundleDeps,
    build_promise_progress_ui_bundle,
)
from .excel_writer_promise_progress_guidance_accuracy import (
    PromiseProgressGuidanceAccuracyDeps,
    build_guidance_accuracy_rows as promise_progress_build_guidance_accuracy_rows,
)
from .excel_writer_promise_progress_orchestrator import (
    PromiseProgressOrchestratorDeps,
    write_promise_progress_ui_v2 as promise_progress_write_promise_progress_ui_v2,
)
from .excel_writer_promise_progress_followthrough import (
    PromiseProgressFollowthroughDeps,
    PromiseProgressFollowthroughModel,
)
from .excel_writer_promise_progress_selection import (
    PromiseProgressSelectionDeps,
    select_promise_progress_rows_for_display,
)
from .excel_writer_promise_progress_rows import (
    PromiseProgressRowsDeps,
    normalize_promise_progress_rows_for_display,
    _dedupe_display_progress_rows as promise_rows_dedupe_display_progress_rows,
    _dedupe_promise_progress_rows as promise_rows_dedupe_promise_progress_rows,
    _display_progress_metric as promise_rows_display_progress_metric,
    _promise_progress_visible_category_rank_local as promise_rows_visible_category_rank_local,
)
from .excel_writer_promise_progress_repairs import (
    GpreProgressTrimDeps,
    PromiseProgressVisibleRepairDeps,
    repair_promise_progress_visible_rows_for_render,
    trim_gpre_final_progress_rows,
)
from .excel_writer_promise_tracker import (
    PromiseTrackerWriterDeps,
    write_promise_tracker_ui_sheet,
)
from .excel_writer_segments import (
    annual_segment_label as ew_annual_segment_label,
    extract_quarter_from_cell as ew_extract_quarter_from_cell,
    extract_segment_line_values as ew_extract_segment_line_values,
    extract_year_from_cell as ew_extract_year_from_cell,
    latest_segment_financials_workbook as ew_latest_segment_financials_workbook,
    parse_annual_segment_data_from_workbook as ew_parse_annual_segment_data_from_workbook,
    parse_quarterly_segment_data_from_workbook as ew_parse_quarterly_segment_data_from_workbook,
    quarterly_segment_label as ew_quarterly_segment_label,
)
from .excel_writer_cached_document_support import (
    CachedDocumentSupport,
    CachedDocumentSupportDeps,
)
from .excel_writer_source_root_support import (
    SourceRootSupport,
    SourceRootSupportDeps,
)
from .excel_writer_sec_cache_support import (
    SecCacheSupport,
    SecCacheSupportDeps,
)
from .excel_writer_sources import (
    build_leverage_audit_doc_index as source_build_leverage_audit_doc_index,
    build_leverage_local_material_index as source_build_leverage_local_material_index,
    build_valuation_filing_docs_by_quarter as source_build_valuation_filing_docs_by_quarter,
    docs_for_valuation_accn as source_docs_for_valuation_accn,
    extract_adj_net_leverage_text_map as source_extract_adj_net_leverage_text_map,
    first_existing_material_dir as source_first_existing_material_dir,
    hist_quarter_whitelist as source_hist_quarter_whitelist,
    infer_cached_doc_quarter as source_infer_cached_doc_quarter,
    infer_q_from_name as source_infer_q_from_name,
    looks_like_leverage_text as source_looks_like_leverage_text,
    normalize_leverage_quarter as source_normalize_leverage_quarter,
    normalize_leverage_text as source_normalize_leverage_text,
    path_cache_key as source_path_cache_key,
    read_cached_doc_raw as source_read_cached_doc_raw,
    read_cached_doc_text as source_read_cached_doc_text,
    resolve_cached_doc_path as source_resolve_cached_doc_path,
    sec_docs_for_accession as source_sec_docs_for_accession,
    slide_text_paths as source_slide_text_paths,
    submission_cache_files as source_submission_cache_files,
    submission_recent_rows as source_submission_recent_rows,
)
from .guidance_lexicon import (
    FORWARD_NOTES_LABEL,
    GUIDANCE_UI_METRIC_PRIORITY,
    classify_metric as glx_classify_metric,
    classify_status as glx_classify_status,
    dedup_text_key as glx_dedup_text_key,
    doc_type_priority as glx_doc_type_priority,
    extract_numeric_patterns as glx_extract_numeric_patterns,
    is_preferred_section as glx_is_preferred_section,
    normalize_text as glx_normalize_text,
    normalize_period as glx_normalize_period,
    score_chunk as glx_score_chunk,
    split_sentences as glx_split_sentences,
)
from .legacy_support import (
    _coerce_next_quarter_end,
    _coerce_prev_quarter_end,
    _extract_balance_sheet_from_html,
    _extract_balance_sheet_from_text,
    _is_quarter_end,
    _path_belongs_to_ticker,
    _prev_quarter_end_from_qend,
    _source_class,
    _source_label,
    _source_method,
    _source_qa,
    _source_tier,
)
from .market_data.service import (
    _build_gpre_proxy_implied_results_bundle as market_build_gpre_proxy_implied_results_bundle,
    _gpre_parse_snapshot_date_like,
    _gpre_footprint_for_quarter,
    _gpre_official_market_weights_for_quarter,
    _gpre_phase_preview_story as market_gpre_phase_preview_story,
    build_current_qtd_simple_crush_snapshot,
    build_gpre_basis_proxy_model,
    build_gpre_overlay_proxy_preview_bundle,
    build_gpre_plant_capacity_history,
    build_gpre_official_proxy_history_series,
    build_gpre_official_proxy_snapshot,
    build_simple_crush_history_series,
    build_prior_quarter_simple_crush_snapshot,
    build_next_quarter_thesis_snapshot,
    fetch_gpre_corn_bids_snapshot,
    load_or_download_gpre_corn_bids_snapshot,
    load_market_export_rows,
    market_input_fingerprint,
    persist_gpre_frozen_thesis_snapshot,
    resolve_gpre_quarter_open_snapshot,
)
from .non_gaap import infer_quarter_end_from_text, parse_adjusted_from_plain_text, strip_html
from .pdf_utils import silence_pdfminer_warnings
from .period_resolver import (
    PickResult,
    _duration_days,
    _filter_unit,
    build_quarter_calendar_from_revenue,
    classify_duration,
    choose_best_tag,
    derive_quarter_from_ytd,
    pick_best_duration,
    pick_best_instant,
    quarter_ends_for_fy,
    self_check_period_logic,
)


def _apply_chart_text_categories(chart_in: Any, *, sheet_name: str, col_idx: int, start_row: int, end_row: int) -> None:
    return ChartTextSupport(
        ChartTextSupportDeps(
            runtime={
                "get_column_letter": get_column_letter,
                "TextAxis": TextAxis,
                "ChartLines": ChartLines,
                "GraphicalProperties": GraphicalProperties,
                "LineProperties": LineProperties,
                "AxDataSource": AxDataSource,
                "StrRef": StrRef,
            }
        )
    ).apply_chart_text_categories(
        chart_in,
        sheet_name=sheet_name,
        col_idx=col_idx,
        start_row=start_row,
        end_row=end_row,
    )
from .pipeline_types import WorkbookInputs
from .quarter_notes_lexicon import (
    compact_snippet as qn_compact_snippet,
    is_complete_signal_text as qn_is_complete_signal_text,
    score_promise_candidate as qn_score_promise_candidate,
    score_quarter_note_candidate as qn_score_quarter_note_candidate,
)
from .operating_drivers_runtime import (
    extract_gpre_45z_accounting_memo,
    OperatingDriversDeps,
    build_operating_drivers_history_rows as runtime_build_operating_drivers_history_rows,
    extract_operating_driver_rows_for_template as runtime_extract_operating_driver_rows_for_template,
    format_operating_driver_delta as runtime_format_operating_driver_delta,
    gpre_canonical_crush_series_for_drivers as runtime_gpre_canonical_crush_series_for_drivers,
    make_driver_row as runtime_make_driver_row,
    merge_driver_rows as runtime_merge_driver_rows,
)
from .sec_xbrl import SecClient, normalize_accession, parse_date
from .signals import build_hidden_value_flags, build_hidden_value_outputs, build_signals_base
from .valuation_precompute_runtime import (
    buyback_execution_scope_text as runtime_buyback_execution_scope_text,
    cap_alloc_unit_mult as runtime_cap_alloc_unit_mult,
    classify_distribution_signal as runtime_classify_distribution_signal,
    extract_cap_alloc_quarter_cash_sentence as runtime_extract_cap_alloc_quarter_cash_sentence,
    extract_cap_alloc_row_cash as runtime_extract_cap_alloc_row_cash,
    extract_valuation_filing_doc_text as runtime_extract_valuation_filing_doc_text,
    has_buyback_execution_table_context as runtime_has_buyback_execution_table_context,
    is_cumulative_buyback_context as runtime_is_cumulative_buyback_context,
    is_debt_repurchase_noise as runtime_is_debt_repurchase_noise,
    parse_cap_alloc_amount as runtime_parse_cap_alloc_amount,
)
from .excel_writer_placeholders import write_empty_sheet_placeholder
from .writer_types import (
    WriterCallbacks,
    WriterContext,
    WriterDerivedData,
    WriterDocumentCache,
    WriterRuntimeData,
)
from .writer_runtime_cache import WriterRuntimeCache
from .writer_qa_policy import latest_quarter_support_gap_severity as writer_qa_latest_quarter_support_gap_severity
from .valuation import valuation_engine, valuation_to_frames


def _net_debt_yoy_flag_label_and_status(delta: Optional[float]) -> Tuple[str, str]:
    """Return a visible net-debt trend label whose wording matches delta sign."""
    if delta is None or pd.isna(delta):
        return "Net debt trend (YoY)", "N/A"
    delta_f = float(delta)
    flat_threshold = 5_000_000.0
    pass_threshold = -50_000_000.0
    if abs(delta_f) <= flat_threshold:
        return "Stable: Net debt roughly flat (YoY)", "WARN"
    if delta_f < 0:
        status = "PASS" if delta_f <= pass_threshold else "WARN"
        return "Green: Net debt decreasing (YoY)", status
    return "Watch: Net debt increasing (YoY)", "FAIL"


def _net_debt_yoy_flag_label_and_status_for_position(
    delta: Optional[float],
    current_net_debt: Optional[float],
) -> Tuple[str, str]:
    """Use net-cash wording when debt increased only because net cash declined."""
    if delta is None or pd.isna(delta):
        return "Net debt trend (YoY)", "N/A"
    try:
        current_f = float(current_net_debt) if current_net_debt is not None and pd.notna(current_net_debt) else None
    except Exception:
        current_f = None
    if current_f is not None and current_f < 0 and float(delta) > 0:
        return "Watch: Net cash decreased YoY", "WARN"
    return _net_debt_yoy_flag_label_and_status(delta)


def _build_compat_state(ctx: WriterContext) -> Dict[str, Any]:
    state = dict(vars(ctx.inputs))
    state.update(
        {
            "out_path": ctx.data.out_path,
            "ticker": ctx.data.ticker,
            "excel_mode": ctx.data.excel_mode,
            "profile_timings": ctx.data.profile_timings,
            "quarter_notes_audit": ctx.data.quarter_notes_audit,
            "enable_operating_drivers_sheet": ctx.data.enable_operating_drivers_sheet,
            "enable_economics_overlay_sheet": ctx.data.enable_economics_overlay_sheet,
            "enable_economics_market_raw_sheet": ctx.data.enable_economics_market_raw_sheet,
            "_driver_inputs_ready": ctx.data.driver_inputs_ready,
            "operating_driver_history_rows": ctx.data.operating_driver_history_rows,
            "economics_market_rows": ctx.data.economics_market_rows,
            "qa_checks": ctx.data.qa_checks,
            "info_log": ctx.data.info_log,
            "data_is_rules_df": ctx.data.data_is_rules_df,
            "writer_timings": ctx.writer_timings,
        }
    )
    state.update(ctx.data.extra_values)
    state.update(ctx.callbacks.as_state_mapping())
    if "_ui_state" in state and "ui_state" not in state:
        state["ui_state"] = state["_ui_state"]
    return state


def _serialize_quarter_note_runtime_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    if isinstance(value, (pd.Timestamp, datetime, date)):
        ts = pd.to_datetime(value, errors="coerce")
        return ts.isoformat() if pd.notna(ts) else str(value)
    if isinstance(value, dict):
        try:
            return json.dumps(value, sort_keys=True, ensure_ascii=True, default=str)
        except Exception:
            return str(value)
    if isinstance(value, (list, tuple, set)):
        try:
            return json.dumps(list(value), ensure_ascii=True, default=str)
        except Exception:
            return str(value)
    return str(value)


def _quarter_note_runtime_signature(item: Dict[str, Any]) -> Tuple[str, ...]:
    return (
        _serialize_quarter_note_runtime_value(item.get("note_id")),
        _serialize_quarter_note_runtime_value(item.get("candidate_type")),
        _serialize_quarter_note_runtime_value(item.get("change_badge")),
        _serialize_quarter_note_runtime_value(item.get("bucket")),
        _serialize_quarter_note_runtime_value(item.get("_metric_display")),
        _serialize_quarter_note_runtime_value(item.get("metric_canon")),
        _serialize_quarter_note_runtime_value(item.get("metric_tag")),
        _serialize_quarter_note_runtime_value(item.get("_split_focus")),
        _serialize_quarter_note_runtime_value(item.get("_render_summary")),
        _serialize_quarter_note_runtime_value(item.get("_pbi_compact_note")),
        _serialize_quarter_note_runtime_value(item.get("target_display")),
        _serialize_quarter_note_runtime_value(item.get("target_period_norm") or item.get("period_key")),
        _serialize_quarter_note_runtime_value(item.get("doc_priority")),
        _serialize_quarter_note_runtime_value(item.get("score")),
        _serialize_quarter_note_runtime_value(item.get("_event_score")),
        _serialize_quarter_note_runtime_value(item.get("theme_key")),
        _serialize_quarter_note_runtime_value(item.get("idea_label")),
        _serialize_quarter_note_runtime_value(item.get("text_full")),
        _serialize_quarter_note_runtime_value(item.get("evidence_snippet")),
        _serialize_quarter_note_runtime_value(item.get("comment_full_text")),
        _serialize_quarter_note_runtime_value(item.get("source") or {}),
    )


def _quarter_note_runtime_qd_token(qd_ref: Optional[date]) -> str:
    q_ts = pd.to_datetime(qd_ref, errors="coerce")
    return q_ts.strftime("%Y-%m-%d") if pd.notna(q_ts) else ""


def _quarter_note_runtime_cache_key(
    cache_name: str,
    item: Dict[str, Any],
    qd_ref: Optional[date],
) -> Tuple[str, str, Tuple[str, ...]]:
    return (
        str(cache_name or ""),
        _quarter_note_runtime_qd_token(qd_ref),
        _quarter_note_runtime_signature(item),
    )


def build_writer_context(inputs: WorkbookInputs) -> WriterContext:
    """Create the run-scoped writer state used by every workbook section.

    This function is the main handoff from pipeline outputs to workbook rendering. It
    copies the normalized `WorkbookInputs`, initializes the workbook object, and sets
    up run-local caches/state that later sheet writers reuse so expensive document,
    valuation, operating-driver, and GPRE economics analysis only happens once per
    export.
    """
    out_path = inputs.out_path
    hist = inputs.hist
    audit = inputs.audit
    needs_review = inputs.needs_review
    debt_tranches = inputs.debt_tranches
    debt_recon = inputs.debt_recon
    adj_metrics = inputs.adj_metrics
    adj_breakdown = inputs.adj_breakdown
    non_gaap_files = inputs.non_gaap_files
    adj_metrics_relaxed = inputs.adj_metrics_relaxed
    adj_breakdown_relaxed = inputs.adj_breakdown_relaxed
    non_gaap_files_relaxed = inputs.non_gaap_files_relaxed
    info_log = inputs.info_log
    tag_coverage = inputs.tag_coverage
    period_checks = inputs.period_checks
    qa_checks = inputs.qa_checks
    bridge_q = inputs.bridge_q
    manifest_df = inputs.manifest_df
    ocr_log = inputs.ocr_log
    qfd_preview = inputs.qfd_preview
    qfd_unused = inputs.qfd_unused
    debt_profile = inputs.debt_profile
    debt_tranches_latest = inputs.debt_tranches_latest
    debt_maturity = inputs.debt_maturity
    debt_credit_notes = inputs.debt_credit_notes
    revolver_df = inputs.revolver_df
    revolver_history = inputs.revolver_history
    debt_buckets = inputs.debt_buckets
    slides_segments = inputs.slides_segments
    slides_debt = inputs.slides_debt
    slides_guidance = inputs.slides_guidance
    quarter_notes = inputs.quarter_notes
    promises = inputs.promises
    promise_progress = inputs.promise_progress
    non_gaap_cred = inputs.non_gaap_cred
    company_overview = inputs.company_overview
    ticker = inputs.ticker
    price = inputs.price
    strictness = inputs.strictness
    excel_mode = inputs.excel_mode
    is_rules = inputs.is_rules
    cache_dir = inputs.cache_dir
    quiet_pdf_warnings = inputs.quiet_pdf_warnings
    rebuild_doc_text_cache = inputs.rebuild_doc_text_cache
    profile_timings = inputs.profile_timings
    quarter_notes_audit = inputs.quarter_notes_audit
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    font_size = 12
    header_size = 13

    wb = Workbook()
    wb.remove(wb.active)
    state: Dict[str, Any] = {}
    ctx_ref: Optional[WriterContext] = None
    # Writer runtime caches are intentionally per-export. They help the many sheet
    # writers share heavy intermediate work inside one workbook build without turning
    # those intermediates into cross-run persisted state.
    runtime_cache = WriterRuntimeCache()
    operating_driver_history_rows: List[Dict[str, Any]] = []
    economics_market_rows: List[Dict[str, Any]] = []
    valuation_style_bundle_cache = runtime_cache.valuation_style_bundle_cache
    valuation_render_bundle_cache = runtime_cache.valuation_render_bundle_cache
    valuation_precompute_bundle_cache = runtime_cache.valuation_precompute_bundle_cache
    valuation_filing_docs_by_quarter_cache = runtime_cache.valuation_filing_docs_by_quarter_cache
    document_cache = WriterDocumentCache()
    frame_view_cache: Dict[Tuple[str, str], pd.DataFrame] = {}
    operating_drivers_runtime = runtime_cache.operating_drivers
    operating_driver_template_index_cache = operating_drivers_runtime.template_index_cache
    operating_driver_bridge_bundle_cache = operating_drivers_runtime.bridge_bundle_cache
    operating_driver_line_index_by_quarter_cache = operating_drivers_runtime.line_index_by_quarter_cache
    operating_driver_flat_line_index_cache = operating_drivers_runtime.flat_line_index_cache
    operating_driver_best_text_cache = operating_drivers_runtime.best_text_cache
    operating_driver_template_rows_cache = operating_drivers_runtime.template_rows_cache
    operating_driver_template_candidate_cache = operating_drivers_runtime.template_candidate_cache
    operating_driver_text_cache = operating_drivers_runtime.text_cache
    profile_slide_signals_cache = operating_drivers_runtime.profile_slide_signals_cache
    profile_slide_signals_by_quarter_cache = operating_drivers_runtime.profile_slide_signals_by_quarter_cache
    operating_driver_45z_guidance_docs_by_quarter_cache = operating_drivers_runtime.guidance_45z_docs_by_quarter_cache
    adj_net_leverage_text_map_cache = runtime_cache.adj_net_leverage_text_map_cache
    leverage_local_material_index_cache = runtime_cache.leverage_local_material_index_cache
    leverage_audit_doc_index_cache = runtime_cache.leverage_audit_doc_index_cache
    promise_progress_ui_bundle_cache = runtime_cache.promise_progress_ui_bundle_cache
    valuation_buyback_auth_source_bundle_cache = runtime_cache.valuation_buyback_auth_source_bundle_cache
    quarter_notes_runtime = runtime_cache.quarter_notes
    valuation_precompute_runtime = runtime_cache.valuation_precompute

    @contextmanager
    def _timed_writer_substage(name: str):
        start = time.perf_counter()
        try:
            yield
        finally:
            _record_writer_substage(name, start)

    def _record_writer_substage(name: str, start: float) -> None:
        dt_s = time.perf_counter() - start
        timings = state.get("writer_timings") if isinstance(state, dict) else None
        if isinstance(timings, dict):
            timings[name] = timings.get(name, 0.0) + dt_s
        if profile_timings:
            print(f"[timing] {name}={dt_s:.2f}s", flush=True)

    def _record_writer_elapsed(name: str, elapsed_s: float) -> None:
        if elapsed_s <= 0:
            return
        timings = state.get("writer_timings") if isinstance(state, dict) else None
        if isinstance(timings, dict):
            timings[name] = timings.get(name, 0.0) + elapsed_s
        if profile_timings:
            print(f"[timing] {name}={elapsed_s:.2f}s", flush=True)

    def _hidden_value_support_runtime() -> Dict[str, Any]:
        return {
            "pd": pd,
            "json": json,
            "wb": wb,
            "font_size": font_size,
            "header_size": header_size,
            "_safe_cell": _safe_cell,
            "_autowidth": _autowidth,
            "Alignment": Alignment,
            "Font": Font,
            "CellIsRule": CellIsRule,
            "PatternFill": PatternFill,
            "Table": Table,
            "TableStyleInfo": TableStyleInfo,
            "get_column_letter": get_column_letter,
            "HiddenValueFlagsSheetInputs": HiddenValueFlagsSheetInputs,
            "write_hidden_value_flags_sheet": write_hidden_value_flags_sheet,
        }

    def _hidden_value_support() -> HiddenValueSupport:
        return HiddenValueSupport(HiddenValueSupportDeps(runtime=_hidden_value_support_runtime()))

    def _build_hidden_value_flags_fallback(flags_audit_df: pd.DataFrame) -> pd.DataFrame:
        return _hidden_value_support().build_hidden_value_flags_fallback(flags_audit_df)
    ui_info_rows: List[Dict[str, Any]] = []

    ticker_roots: List[Path] = []
    repo_root = Path(__file__).resolve().parents[2]

    def _source_root_support_runtime() -> Dict[str, Any]:
        try:
            profile_ticker_runtime = profile_ticker
        except NameError:
            profile_ticker_runtime = ticker
        return {
            "Path": Path,
            "re": re,
            "ticker": ticker,
            "profile_ticker": profile_ticker_runtime,
            "repo_root": repo_root,
            "out_path": out_path,
            "cache_dir": cache_dir,
            "manifest_df": manifest_df,
            "canonical_shared_cache_root": canonical_shared_cache_root,
            "ticker_cache_candidates": ticker_cache_candidates,
            "ticker_cache_roots_from_base_dir": ticker_cache_roots_from_base_dir,
            "_path_belongs_to_ticker": _path_belongs_to_ticker,
        }

    def _get_source_root_support() -> SourceRootSupport:
        return SourceRootSupport(SourceRootSupportDeps(runtime=_source_root_support_runtime()))

    def _path_within_scope(path_in: Any, root_in: Any) -> bool:
        return _get_source_root_support().path_within_scope(path_in, root_in)

    def _company_material_roots() -> List[Path]:
        return _get_source_root_support().company_material_roots(ticker_roots)

    material_roots = _company_material_roots()
    company_profile = get_company_profile(ticker)
    profile_ticker = str(getattr(company_profile, "ticker", "") or ticker or "").strip().upper()
    is_pbi_profile = profile_ticker == "PBI"
    is_gpre_profile = profile_ticker == "GPRE"
    is_anf_profile = profile_ticker == "ANF"

    def _is_repo_profile_cache_path(path_in: Any) -> bool:
        return _get_source_root_support().is_repo_profile_cache_path(path_in)

    def _allow_repo_profile_cache_fallback() -> bool:
        return _get_source_root_support().allow_repo_profile_cache_fallback()
    enable_operating_drivers_sheet = bool(
        getattr(company_profile, "enable_operating_drivers_sheet", False)
    )
    enable_economics_overlay_sheet = bool(
        getattr(company_profile, "enable_economics_overlay_sheet", False)
    )
    enable_economics_market_raw_sheet = bool(
        getattr(company_profile, "enable_economics_market_raw_sheet", False)
    )
    enable_quarterly_segment_block = bool(
        getattr(company_profile, "enable_quarterly_segment_block", False)
    )
    quarterly_segment_labels = tuple(
        getattr(company_profile, "quarterly_segment_labels", tuple()) or tuple()
    )
    enable_annual_segment_block = bool(
        getattr(company_profile, "enable_annual_segment_block", True)
    )
    annual_segment_labels = tuple(getattr(company_profile, "annual_segment_labels", tuple()) or tuple())
    annual_segment_alias_patterns = tuple(
        getattr(company_profile, "annual_segment_alias_patterns", tuple()) or tuple()
    )
    pbi_summary_description_fallback = str(
        getattr(company_profile, "summary_description_fallback", "") or ""
    ).strip()
    pbi_summary_dependency_fallbacks = tuple(
        getattr(company_profile, "summary_dependency_fallbacks", tuple()) or tuple()
    )
    pbi_summary_wrong_thesis_fallbacks = tuple(
        getattr(company_profile, "summary_wrong_thesis_fallbacks", tuple()) or tuple()
    )
    profile_signal_runtime: Dict[str, Any] = {**globals(), **locals()}
    profile_signal_support: Optional[ProfileSignalSupport] = None

    def _refresh_profile_signal_runtime(runtime_update: Optional[Mapping[str, Any]] = None) -> None:
        if runtime_update is not None:
            profile_signal_runtime.update(runtime_update)

    def _get_profile_signal_support() -> ProfileSignalSupport:
        nonlocal profile_signal_support
        if profile_signal_support is None:
            profile_signal_support = ProfileSignalSupport(
                ProfileSignalSupportDeps(runtime=profile_signal_runtime)
            )
        return profile_signal_support

    def _is_preferred_narrative_source(source_type_in: Any) -> bool:
        return _get_profile_signal_support().is_preferred_narrative_source(source_type_in)

    def _looks_pbi_fragment_text(text_in: Any) -> bool:
        return _get_profile_signal_support().looks_pbi_fragment_text(text_in)

    def _is_pbi_clean_sentence(text_in: Any) -> bool:
        return _get_profile_signal_support().is_pbi_clean_sentence(text_in)

    def _pbi_target_display_ok(text_in: Any) -> bool:
        return _get_profile_signal_support().pbi_target_display_ok(text_in)

    def _profile_signal_helper(name: str, *args: Any, **kwargs: Any) -> Any:
        return _get_profile_signal_support().call_helper(name, *args, **kwargs)

    def _classify_pbi_metric_label(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_classify_pbi_metric_label", *args, **kwargs)

    def _extract_pbi_guidance_targets_multi(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_extract_pbi_guidance_targets_multi", *args, **kwargs)

    def _extract_pbi_target_display(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_extract_pbi_target_display", *args, **kwargs)

    def _first_existing_material_dir(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_first_existing_material_dir", *args, **kwargs)

    def _parse_quarter_from_filename(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_parse_quarter_from_filename", *args, **kwargs)

    def _pbi_reported_fcf_payload_for_qd(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_reported_fcf_payload_for_qd", *args, **kwargs)

    def _ensure_terminal_period(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_ensure_terminal_period", *args, **kwargs)

    def _parse_gpre_crush_margin_pair_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_parse_gpre_crush_margin_pair_local", *args, **kwargs)

    def _collapse_repeated_leading_ngram_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_collapse_repeated_leading_ngram_local", *args, **kwargs)

    def _dedupe_canonical_text_parts_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_dedupe_canonical_text_parts_local", *args, **kwargs)

    def _pbi_guidance_period_label_from_text(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_guidance_period_label_from_text", *args, **kwargs)

    def _pbi_structured_guidance_items_for_qd(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_structured_guidance_items_for_qd", *args, **kwargs)

    def _period_label_to_norm(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_period_label_to_norm", *args, **kwargs)

    def _pbi_repair_guidance_period_meta(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_repair_guidance_period_meta", *args, **kwargs)

    def _lookup_pbi_structured_guidance_target(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_lookup_pbi_structured_guidance_target", *args, **kwargs)

    def _lookup_pbi_structured_progress_hint(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_lookup_pbi_structured_progress_hint", *args, **kwargs)

    def _pbi_structured_strategy_items_for_qd(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_structured_strategy_items_for_qd", *args, **kwargs)

    def _parse_quarter_from_follow_text(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_parse_quarter_from_follow_text", *args, **kwargs)

    def _profile_slide_metric(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_profile_slide_metric", *args, **kwargs)

    def _extract_money_targets_for_display(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_extract_money_targets_for_display", *args, **kwargs)

    def _extract_45z_2026_target_candidates(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_extract_45z_2026_target_candidates", *args, **kwargs)

    def _strong_45z_2026_target_display(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_strong_45z_2026_target_display", *args, **kwargs)

    def _coerce_amount_with_unit_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_coerce_amount_with_unit_local", *args, **kwargs)

    def _fmt_short_money_value_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_fmt_short_money_value_local", *args, **kwargs)

    def _is_45z_crush_margin_support_only(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_is_45z_crush_margin_support_only", *args, **kwargs)

    def _extract_45z_monetization_target_display(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_extract_45z_monetization_target_display", *args, **kwargs)

    def _extract_45z_realized_progress_text(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_extract_45z_realized_progress_text", *args, **kwargs)

    def _slide_signal_noise(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_slide_signal_noise", *args, **kwargs)

    def _text_fragment_penalty(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_text_fragment_penalty", *args, **kwargs)

    def _clean_target_bonus(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_clean_target_bonus", *args, **kwargs)

    def _source_rank(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_source_rank", *args, **kwargs)

    def _management_theme_key(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_management_theme_key", *args, **kwargs)

    def _split_target_group_key(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_split_target_group_key", *args, **kwargs)

    def _split_target_qend(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_split_target_qend", *args, **kwargs)

    def _split_target_family_key(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_split_target_family_key", *args, **kwargs)

    def _derive_split_target_meta(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_derive_split_target_meta", *args, **kwargs)

    def _split_target_metric_display(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_split_target_metric_display", *args, **kwargs)

    def _gpre_normalize_metric_label(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_gpre_normalize_metric_label", *args, **kwargs)

    def _gpre_clean_visible_promise_metric(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_gpre_clean_visible_promise_metric", *args, **kwargs)

    def _gpre_bad_visible_promise_reason(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_gpre_bad_visible_promise_reason", *args, **kwargs)

    def _split_target_scope_token(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_split_target_scope_token", *args, **kwargs)

    def _split_target_scope_is_broad(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_split_target_scope_is_broad", *args, **kwargs)

    def _target_period_is_closed(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_target_period_is_closed", *args, **kwargs)

    def _infer_target_period(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_infer_target_period", *args, **kwargs)

    def _nearest_amount_for_pattern(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_nearest_amount_for_pattern", *args, **kwargs)

    def _infer_target_structure(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_infer_target_structure", *args, **kwargs)

    def _split_target_identity_key(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_split_target_identity_key", *args, **kwargs)

    def _candidate_quality_key(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_candidate_quality_key", *args, **kwargs)
    def _pbi_forced_release_backed_fcf_note_for_qd(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_forced_release_backed_fcf_note_for_qd", *args, **kwargs)

    def _pbi_guidance_period_for_text(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_guidance_period_for_text", *args, **kwargs)

    def _fmt_pbi_million_amount(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_fmt_pbi_million_amount", *args, **kwargs)

    def _pbi_guidance_table_target_display(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_guidance_table_target_display", *args, **kwargs)

    def _format_directional_fcf_summary_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_format_directional_fcf_summary_local", *args, **kwargs)

    def _format_directional_from_prior_summary_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_format_directional_from_prior_summary_local", *args, **kwargs)

    def _parse_signed_money_token_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_parse_signed_money_token_local", *args, **kwargs)

    def _pbi_guidance_self_contained_summary(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_guidance_self_contained_summary", *args, **kwargs)

    def _pbi_guidance_compact_note(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_guidance_compact_note", *args, **kwargs)

    def _pbi_compact_guidance_note(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_compact_guidance_note", *args, **kwargs)

    def _pbi_guidance_sentence(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_guidance_sentence", *args, **kwargs)

    def _pbi_period_label_from_norm(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_period_label_from_norm", *args, **kwargs)

    def _pbi_guidance_period_norm_is_reasonable(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_guidance_period_norm_is_reasonable", *args, **kwargs)

    def _pbi_default_guidance_period_label(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_default_guidance_period_label", *args, **kwargs)

    def _pbi_parse_money_amount(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_parse_money_amount", *args, **kwargs)

    def _pbi_parse_money_range(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_parse_money_range", *args, **kwargs)

    def _pbi_strategy_target_display(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_strategy_target_display", *args, **kwargs)

    def _pbi_structured_note_rows_for_qd(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_pbi_structured_note_rows_for_qd", *args, **kwargs)

    def _profile_slide_category(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_profile_slide_category", *args, **kwargs)

    def _fmt_short_money_value_with_parens_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_fmt_short_money_value_with_parens_local", *args, **kwargs)

    def _profile_slide_target_display(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_profile_slide_target_display", *args, **kwargs)

    def _slide_line_is_noise(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_slide_line_is_noise", *args, **kwargs)

    def _slide_line_is_heading(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_slide_line_is_heading", *args, **kwargs)

    def _extract_profile_signal_texts(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_extract_profile_signal_texts", *args, **kwargs)

    def _clean_profile_signal_text(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_clean_profile_signal_text", *args, **kwargs)

    def _split_target_scope_key(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_split_target_scope_key", *args, **kwargs)

    def _clean_split_target_scope_label(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_clean_split_target_scope_label", *args, **kwargs)

    def _is_time_scope_label(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_is_time_scope_label", *args, **kwargs)

    def _extract_named_split_target_scope(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_extract_named_split_target_scope", *args, **kwargs)

    def _gpre_clean_visible_note_metric(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_gpre_clean_visible_note_metric", *args, **kwargs)

    def _quarter_bounds(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_quarter_bounds", *args, **kwargs)

    def _half_bounds(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_half_bounds", *args, **kwargs)

    def _extract_target_amounts_with_spans(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_extract_target_amounts_with_spans", *args, **kwargs)

    def _slide_excerpt(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_slide_excerpt", *args, **kwargs)

    def _profile_slide_signals_cache_path(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_profile_slide_signals_cache_path", *args, **kwargs)

    def _profile_slide_signals_signature(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_profile_slide_signals_signature", *args, **kwargs)

    def _profile_slide_signal_jsonable(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_profile_slide_signal_jsonable", *args, **kwargs)

    def _read_profile_slide_signals_cache(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_read_profile_slide_signals_cache", *args, **kwargs)

    def _write_profile_slide_signals_cache(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_write_profile_slide_signals_cache", *args, **kwargs)

    def _quarter_matches_source_event_window_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_quarter_matches_source_event_window_local", *args, **kwargs)

    def _gpre_official_risk_management_note_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_gpre_official_risk_management_note_local", *args, **kwargs)

    def _gpre_transcript_risk_management_note_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_gpre_transcript_risk_management_note_local", *args, **kwargs)

    def _gpre_official_45z_realized_note_local(*args: Any, **kwargs: Any) -> Any:
        return _profile_signal_helper("_gpre_official_45z_realized_note_local", *args, **kwargs)
    def _load_profile_slide_signals() -> List[Dict[str, Any]]:
        nonlocal profile_slide_signals_cache
        rows = _get_profile_signal_support().load_profile_slide_signals()
        profile_slide_signals_cache = _get_profile_signal_support().profile_slide_signals_cache
        profile_signal_runtime["profile_slide_signals_cache"] = profile_slide_signals_cache
        return rows

    def _load_profile_slide_signals_by_quarter() -> Dict[date, List[Dict[str, Any]]]:
        nonlocal profile_slide_signals_by_quarter_cache
        grouped = _get_profile_signal_support().load_profile_slide_signals_by_quarter()
        profile_slide_signals_by_quarter_cache = _get_profile_signal_support().profile_slide_signals_by_quarter_cache
        profile_signal_runtime["profile_slide_signals_by_quarter_cache"] = profile_slide_signals_by_quarter_cache
        return grouped

    def _profile_slide_signals_for_quarter(qd: date) -> List[Dict[str, Any]]:
        return _get_profile_signal_support().profile_slide_signals_for_quarter(qd)

    def _local_slide_driver_fallback(qd: date, driver_kind: str) -> str:
        return _get_profile_signal_support().local_slide_driver_fallback(qd, driver_kind)

    def _pbi_slide_pages_for_qd(qd: Optional[date]) -> List[Dict[str, Any]]:
        return _get_profile_signal_support().pbi_slide_pages_for_qd(qd)

    def _local_slide_45z_realized_text(qd: date) -> str:
        return _get_profile_signal_support().local_slide_45z_realized_text(qd)

    _pbi_note_theme_re = _get_profile_signal_support().pbi_note_theme_re
    _pbi_promise_theme_re = _get_profile_signal_support().pbi_promise_theme_re

    def _cache_roots() -> List[Path]:
        return _get_source_root_support().cache_roots(material_roots)

    cache_roots = _cache_roots()
    cache_root = next((p for p in cache_roots if p.exists()), Path(__file__).resolve().parents[2] / "sec_cache")
    pdf_text_cache_root = Path(cache_dir) if cache_dir is not None else cache_root
    local_balance_sheet_support_state: Dict[str, Any] = {
        "payload_cache": {},
        "file_index_cache": None,
        "records_by_quarter_cache": None,
        "quarter_cache": {},
        "payload_by_path_cache": {},
    }

    def _read_material_text(path_in: Path) -> str:
        return _read_cached_doc_text(path_in)

    def _local_balance_sheet_support_runtime() -> Dict[str, Any]:
        return {
            "pd": pd,
            "ticker": ticker,
            "material_roots": material_roots,
            "ticker_roots": ticker_roots,
            "local_balance_sheet_support_state": local_balance_sheet_support_state,
            "_path_belongs_to_ticker": _path_belongs_to_ticker,
            "_path_cache_key": _path_cache_key,
            "_parse_quarter_from_filename": _parse_quarter_from_filename,
            "_parse_quarter_from_follow_text": _parse_quarter_from_follow_text,
            "infer_quarter_end_from_text": infer_quarter_end_from_text,
            "_extract_balance_sheet_from_html": _extract_balance_sheet_from_html,
            "_extract_balance_sheet_from_text": _extract_balance_sheet_from_text,
            "_read_material_text": _read_material_text,
            "_timed_writer_substage": _timed_writer_substage,
        }

    def _get_local_balance_sheet_support() -> LocalBalanceSheetSupport:
        return LocalBalanceSheetSupport(
            LocalBalanceSheetSupportDeps(runtime=_local_balance_sheet_support_runtime())
        )

    def _shared_financial_statement_files() -> List[Path]:
        return _get_local_balance_sheet_support().shared_financial_statement_files()

    def _shared_local_balance_sheet_quarter(rec: Dict[str, Any]) -> Optional[date]:
        return _get_local_balance_sheet_support().shared_local_balance_sheet_quarter(rec)

    def _shared_local_balance_sheet_records_by_quarter() -> Dict[date, List[Dict[str, Any]]]:
        return _get_local_balance_sheet_support().shared_local_balance_sheet_records_by_quarter()

    def _shared_local_balance_sheet_payload_for_record(rec: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        return _get_local_balance_sheet_support().shared_local_balance_sheet_payload_for_record(rec)

    def _shared_load_local_balance_sheet_detail_payloads(
        target_quarters: Optional[set[date]] = None,
    ) -> Dict[date, Dict[str, Any]]:
        return _get_local_balance_sheet_support().shared_load_local_balance_sheet_detail_payloads(target_quarters)

    def _carry_forward_low_change_series(
        src_map: Dict[pd.Timestamp, Optional[float]],
        q_series: List[Any],
        *,
        max_gap_quarters: int = 4,
        rel_tol: float = 1e-4,
        abs_tol: float = 1_000.0,
    ) -> Dict[pd.Timestamp, Optional[float]]:
        return _get_local_balance_sheet_support().carry_forward_low_change_series(
            src_map,
            q_series,
            max_gap_quarters=max_gap_quarters,
            rel_tol=rel_tol,
            abs_tol=abs_tol,
        )

    if slides_segments is None:
        slides_segments = pd.DataFrame()
    if slides_debt is None:
        slides_debt = pd.DataFrame()
    if slides_guidance is None:
        slides_guidance = pd.DataFrame()
    if quarter_notes is None:
        quarter_notes = pd.DataFrame()
    if promises is None:
        promises = pd.DataFrame()
    if promise_progress is None:
        promise_progress = pd.DataFrame()
    if non_gaap_cred is None:
        non_gaap_cred = pd.DataFrame()
    if company_overview is None:
        company_overview = {
            "what_it_does": "N/A",
            "what_it_does_source": "Source: N/A (company overview not available)",
            "current_strategic_context": "N/A",
            "current_strategic_context_source": "Source: N/A (current strategic context not available)",
            "key_advantage": "N/A",
            "key_advantage_source": "Source: N/A (company overview not available)",
            "revenue_streams": [],
            "revenue_streams_source": "Source: N/A (company overview not available)",
            "asof_fy_end": None,
        }
    def _bank_metrics_enabled() -> bool:
        if bool(getattr(company_profile, "has_bank", False)):
            return True
        if hist is None or hist.empty:
            return False
        bank_metrics = ["bank_deposits", "bank_finance_receivables", "bank_net_funding"]
        h_chk = hist.copy()
        if "quarter" in h_chk.columns:
            h_chk["quarter"] = pd.to_datetime(h_chk["quarter"], errors="coerce")
            h_chk = h_chk[h_chk["quarter"].notna()].sort_values("quarter").tail(8)
        qualifying = 0
        for metric in bank_metrics:
            if metric not in h_chk.columns:
                continue
            vals = pd.to_numeric(h_chk[metric], errors="coerce")
            if int(vals.notna().sum()) >= 4:
                qualifying += 1
        return qualifying >= 2

    bank_metrics_enabled = _bank_metrics_enabled()

    def _build_numeric_quarter_map(df_in: Optional[pd.DataFrame], col_name: str) -> Dict[date, float]:
        if df_in is None or getattr(df_in, "empty", True) or col_name not in df_in.columns:
            return {}
        try:
            df_local = df_in.copy()
            df_local["quarter"] = pd.to_datetime(df_local.get("quarter"), errors="coerce").dt.date
            df_local[col_name] = pd.to_numeric(df_local.get(col_name), errors="coerce")
            df_local = df_local[df_local["quarter"].notna() & df_local[col_name].notna()]
            return {q: float(v) for q, v in zip(df_local["quarter"], df_local[col_name])}
        except Exception:
            return {}

    _pbi_hist_buybacks_cash_map = _build_numeric_quarter_map(hist, "buybacks_cash") if is_pbi_profile else {}
    _pbi_hist_debt_repayment_map = _build_numeric_quarter_map(hist, "debt_repayment") if is_pbi_profile else {}
    _pbi_adj_fcf_map = _build_numeric_quarter_map(adj_metrics, "adj_fcf") if is_pbi_profile else {}
    _pbi_revolver_availability_map = _build_numeric_quarter_map(revolver_history, "revolver_availability") if is_pbi_profile else {}

    def _prev_same_quarter_year(qd: Optional[date], value_map: Dict[date, float]) -> Optional[float]:
        if not qd or not value_map:
            return None
        try:
            candidate = qd.replace(year=qd.year - 1)
        except Exception:
            return None
        return value_map.get(candidate)

    def _prev_available_quarter(qd: Optional[date], value_map: Dict[date, float]) -> Optional[Tuple[date, float]]:
        if not qd or not value_map:
            return None
        prior_keys = [qq for qq in value_map.keys() if isinstance(qq, date) and qq < qd]
        if not prior_keys:
            return None
        best_q = max(prior_keys)
        return best_q, float(value_map.get(best_q))

    def _safe_cell(val: Any) -> Any:
        if isinstance(val, str):
            val = ILLEGAL_CHARACTERS_RE.sub("", val)
            if val.startswith("="):
                return "'" + val
        return val

    def _safe_cell_or_none(val: Any) -> Any:
        """Normalize pandas missing values and Excel-unsafe strings before writing.

        The workbook writer pushes many large diagnostic/raw-data frames through
        openpyxl. Keeping this tiny conversion centralized avoids slower
        DataFrame.iterrows access patterns while preserving the existing cell
        safety contract for formulas-as-text and illegal XML characters.
        """
        if val is None:
            return None
        try:
            if pd.isna(val):
                return None
        except Exception:
            pass
        return _safe_cell(val)

    def _excel_safe_text_local(text_in: Any, *, max_len: Optional[int] = None) -> str:
        txt = str(text_in or "")
        if not txt:
            return ""
        txt = txt.replace("\r\n", "\n").replace("\r", "\n")
        txt = ILLEGAL_CHARACTERS_RE.sub("", txt)
        if max_len is not None:
            txt = txt[:max_len]
        return txt

    def _excel_safe_comment_text_local(text_in: Any, *, max_len: int = 32000) -> str:
        txt = _excel_safe_text_local(text_in)
        if not txt:
            return ""
        txt = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", txt)
        txt = txt.replace("\r\n", "\n").replace("\r", "\n")
        if max_len is not None:
            txt = txt[:max_len]
        return txt

    def _comment_local(text_in: Any, author: str = "pipeline") -> Optional[Comment]:
        txt = _excel_safe_comment_text_local(text_in)
        if not txt:
            return None
        return Comment(txt, author)

    def _set_cell_comment_local(cell: Any, text_in: Any, author: str = "pipeline") -> None:
        comment = _comment_local(text_in, author=author)
        if comment is None:
            return
        cell.comment = comment

    def _updated_font(existing_font: Any, **changes: Any) -> Any:
        font_obj = copy(existing_font)
        for attr, value in changes.items():
            setattr(font_obj, attr, value)
        return font_obj

    def _apply_hyperlink_look(cell, target: str) -> None:
        if not target:
            return
        cell.hyperlink = target
        cell.font = _updated_font(cell.font, color="0563C1", underline="single")

    # Backward-compatible alias for older writer helpers.
    def _apply_hyperlink(cell, target: str) -> None:
        _apply_hyperlink_look(cell, target)

    def _normalize_accn_local(val: Any) -> str:
        return re.sub(r"[^0-9]", "", str(val or ""))

    sec_cache_support: Optional[SecCacheSupport] = None

    def _sec_cache_support_runtime() -> Dict[str, Any]:
        return {
            "Path": Path,
            "re": re,
            "date": date,
            "ticker": ticker,
            "cache_dir": cache_dir,
            "document_cache": document_cache,
            "ticker_cache_candidates": ticker_cache_candidates,
            "_parse_quarter_from_filename": _parse_quarter_from_filename,
            "_parse_quarter_from_follow_text": _parse_quarter_from_follow_text,
            "_read_cached_doc_text": lambda path_in: _read_cached_doc_text(path_in),
            "infer_quarter_end_from_text": infer_quarter_end_from_text,
        }

    def _get_sec_cache_support() -> SecCacheSupport:
        nonlocal sec_cache_support
        if sec_cache_support is None:
            sec_cache_support = SecCacheSupport(
                SecCacheSupportDeps(runtime=_sec_cache_support_runtime())
            )
        return sec_cache_support

    def _sec_cache_roots_local() -> List[Path]:
        return _get_sec_cache_support().sec_cache_roots_local()

    def _sec_cache_doc_paths_local(root: Path) -> List[Path]:
        return _get_sec_cache_support().sec_cache_doc_paths_local(root)

    def _sec_cache_html_paths_local(root: Path) -> List[Path]:
        return _get_sec_cache_support().sec_cache_html_paths_local(root)

    def _sec_cache_docs_for_token_local(root: Path, token: str) -> List[Path]:
        return _get_sec_cache_support().sec_cache_docs_for_token_local(root, token)

    def _sec_cache_html_paths_for_token_local(root: Path, token: str) -> List[Path]:
        return _get_sec_cache_support().sec_cache_html_paths_for_token_local(root, token)

    def _infer_doc_quarter_local(path_in: Any, raw_text: Any = "") -> Optional[date]:
        return _get_sec_cache_support().infer_doc_quarter_local(path_in, raw_text)

    debt_convertible_support: Optional[DebtConvertibleEnrichmentSupport] = None

    def _htmlish_to_text(txt_in: str) -> str:
        txt = html.unescape(str(txt_in or ""))
        txt = re.sub(r"<br\s*/?>", "\n", txt, flags=re.I)
        txt = re.sub(r"</?(?:div|p|tr|li|td|th|table|span|font|b|strong|em|u)[^>]*>", " ", txt, flags=re.I)
        txt = re.sub(r"<[^>]+>", " ", txt)
        txt = txt.replace("\xa0", " ")
        return glx_normalize_text(txt)

    def _safe_text_value(v: Any) -> str:
        try:
            if pd.isna(v):
                return ""
        except Exception:
            pass
        return str(v or "").strip()

    def _get_debt_convertible_enrichment_support() -> DebtConvertibleEnrichmentSupport:
        nonlocal debt_convertible_support
        if debt_convertible_support is None:
            debt_convertible_support = DebtConvertibleEnrichmentSupport(
                DebtConvertibleEnrichmentDeps(
                    ticker=ticker,
                    cache_dir=cache_dir,
                    ticker_roots=tuple(ticker_roots),
                    document_cache=document_cache,
                    context_helpers={
                        "_normalize_accn_local": _normalize_accn_local,
                        "_sec_cache_roots_local": _sec_cache_roots_local,
                        "_sec_cache_doc_paths_local": _sec_cache_doc_paths_local,
                        "_infer_doc_quarter_local": _infer_doc_quarter_local,
                        "_path_belongs_to_ticker": _path_belongs_to_ticker,
                        "coerce_number": coerce_number,
                        "glx_normalize_text": glx_normalize_text,
                    },
                )
            )
        return debt_convertible_support

    def _enrich_latest_debt_convertibles(df_in: pd.DataFrame) -> pd.DataFrame:
        return _get_debt_convertible_enrichment_support().enrich_latest_debt_convertibles(df_in)

    def _build_operating_driver_rows() -> List[Dict[str, Any]]:
        return _get_operating_drivers_support().build_operating_driver_rows()

    valuation_bridge_support: Optional[ValuationBridgeSupport] = None

    def _valuation_bridge_support_runtime() -> Dict[str, Any]:
        return {
            "pd": pd,
            "re": re,
            "Path": Path,
            "ticker": ticker,
            "hist": hist,
            "adj_metrics": adj_metrics,
            "_adj_metrics_view": _adj_metrics_view,
            "_operating_driver_financial_statement_files": _operating_driver_financial_statement_files,
            "_operating_driver_follow_source_dirs": _operating_driver_follow_source_dirs,
            "_read_operating_driver_text": _read_operating_driver_text,
            "glx_normalize_text": glx_normalize_text,
            "qn_compact_snippet": qn_compact_snippet,
            "_source_rank": _source_rank,
        }

    def _get_valuation_bridge_support() -> ValuationBridgeSupport:
        nonlocal valuation_bridge_support
        runtime = _valuation_bridge_support_runtime()
        if valuation_bridge_support is None:
            valuation_bridge_support = ValuationBridgeSupport(
                ValuationBridgeSupportDeps(runtime=runtime)
            )
        else:
            valuation_bridge_support.runtime.update(runtime)
        return valuation_bridge_support

    def _load_bridge_fy_adjusted_ebitda_records() -> List[Dict[str, Any]]:
        return _get_valuation_bridge_support().load_bridge_fy_adjusted_ebitda_records()

    def _resolve_thesis_fy_base() -> Dict[str, Any]:
        return _get_valuation_bridge_support().resolve_thesis_fy_base()

    debt_tranches_latest = _enrich_latest_debt_convertibles(debt_tranches_latest)

    def _write_sheet(name: str, df: pd.DataFrame) -> None:
        ws = wb.create_sheet(name)
        if str(name or "") == "History_Q" and isinstance(df, pd.DataFrame) and not df.empty:
            df = _augment_history_q_frame_for_writer(df, ticker=ticker, fiscal_profile=company_profile)
        if df is None or df.empty:
            write_empty_sheet_placeholder(ws)
            return
        header_vals = []
        for c in df.columns:
            if isinstance(c, (pd.Timestamp, datetime, date)):
                header_vals.append(pd.to_datetime(c).date())
            else:
                header_vals.append(c)
        ws.append(header_vals)
        # itertuples avoids allocating a Series for every row. This matters for
        # DATA_Facts_Long and other large raw-data sheets, where row-by-row
        # pandas indexing can dominate workbook write time.
        for row_values in df.itertuples(index=False, name=None):
            ws.append([_safe_cell_or_none(value) for value in row_values])
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.font = Font(bold=True, size=header_size)
            c.alignment = Alignment(vertical="center")
        _autowidth(ws, len(df.columns))
        ws.sheet_format.defaultRowHeight = 18
        ws.sheet_view.zoomScale = 110
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                size = header_size if cell.row == 1 else font_size
                cell.font = _updated_font(cell.font, size=size, bold=cell.font.b)
        try:
            # Skip table if headers are not plain strings (Excel tables require string headers)
            headers = [c.value for c in ws[1]]
            if any(h is None or isinstance(h, (datetime, date)) or not isinstance(h, str) for h in headers):
                raise ValueError("Non-string headers; skip table")
            if len(headers) != len(set(headers)):
                raise ValueError("Duplicate headers; skip table")
            ref = f"A1:{get_column_letter(len(df.columns))}{ws.max_row}"
            t = Table(displayName=name.replace(" ", "").replace("-", ""), ref=ref)
            t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(t)
        except Exception:
            pass

    def _derivative_oci_bridge_render_runtime() -> Dict[str, Any]:
        return {
            "wb": wb,
            "_write_sheet": _write_sheet,
            "_safe_cell": _safe_cell,
            "_get_analysis_sheet_style_bundle": _get_analysis_sheet_style_bundle,
            "font_size": font_size,
            "header_size": header_size,
            "ctx_ref": ctx_ref,
            "operating_driver_history_rows": operating_driver_history_rows,
        }

    def _write_derivative_oci_bridge_sheet(
        bridge_df: pd.DataFrame,
        exposure_df: Optional[pd.DataFrame] = None,
    ) -> None:
        write_derivative_oci_bridge_sheet(
            DerivativeOciBridgeRenderDeps(runtime=_derivative_oci_bridge_render_runtime()),
            bridge_df,
            exposure_df,
        )

    def _write_derivative_crush_tests_sheet(tables: Dict[str, pd.DataFrame]) -> None:
        write_derivative_crush_tests_sheet(
            DerivativeOciBridgeRenderDeps(runtime=_derivative_oci_bridge_render_runtime()),
            tables,
        )

    def _write_flags_sheet(name: str, df: pd.DataFrame) -> None:
        return _hidden_value_support().write_flags_sheet(name, df)

    def _autowidth(ws, n_cols: int) -> None:
        for i in range(1, n_cols + 1):
            width = max(14, ws.column_dimensions[get_column_letter(i)].width or 14)
            if i == 1:
                width = max(width, 26)
            ws.column_dimensions[get_column_letter(i)].width = width

    def _norm_header_key(name: Any) -> str:
        s = str(name or "").strip().lower()
        s = re.sub(r"[^a-z0-9]+", "_", s)
        s = re.sub(r"_+", "_", s).strip("_")
        return s

    def _resolve_col(df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
        if df is None or df.empty:
            return None
        norm_map: Dict[str, str] = {}
        for c in df.columns:
            k = _norm_header_key(c)
            if k and k not in norm_map:
                norm_map[k] = c
        for alias in aliases:
            k = _norm_header_key(alias)
            if k in norm_map:
                return norm_map[k]
        return None

    def _cached_source_view(
        source_name: str,
        df_in: Any,
        *,
        quarter_aliases: List[str],
        quarter_mode: str = "timestamp",
    ) -> pd.DataFrame:
        cache_key = (source_name, quarter_mode)
        cached = frame_view_cache.get(cache_key)
        if cached is not None:
            return cached
        df = df_in.copy() if isinstance(df_in, pd.DataFrame) else pd.DataFrame()
        q_col = _resolve_col(df, quarter_aliases)
        if q_col is not None:
            q_series = pd.to_datetime(df[q_col], errors="coerce")
            if quarter_mode == "date":
                df["_quarter"] = q_series.dt.date
            elif quarter_mode == "quarter_end_date":
                df["_quarter"] = q_series.dt.to_period("Q").dt.end_time.dt.date
            else:
                df["_quarter"] = q_series
        frame_view_cache[cache_key] = df
        return df

    def _hist_view(*, quarter_mode: str = "timestamp") -> pd.DataFrame:
        return _cached_source_view("hist", hist, quarter_aliases=["quarter"], quarter_mode=quarter_mode)

    def _adj_metrics_view(*, quarter_mode: str = "timestamp") -> pd.DataFrame:
        return _cached_source_view("adj_metrics", adj_metrics, quarter_aliases=["quarter"], quarter_mode=quarter_mode)

    def _quarter_notes_view(*, quarter_mode: str = "timestamp") -> pd.DataFrame:
        return _cached_source_view(
            "quarter_notes",
            quarter_notes,
            quarter_aliases=["quarter", "quarter_end", "as_of_quarter"],
            quarter_mode=quarter_mode,
        )

    def _promises_view(*, quarter_mode: str = "timestamp") -> pd.DataFrame:
        return _cached_source_view(
            "promises",
            promises,
            quarter_aliases=["last_seen_quarter", "quarter", "created_quarter", "first_seen_quarter"],
            quarter_mode=quarter_mode,
        )

    def _audit_view(*, quarter_mode: str = "timestamp") -> pd.DataFrame:
        return _cached_source_view(
            "audit",
            audit,
            quarter_aliases=["quarter", "quarter_end", "period_end"],
            quarter_mode=quarter_mode,
        )

    def _qend_date(x: Any) -> Optional[date]:
        t = pd.to_datetime(x, errors="coerce")
        if pd.isna(t):
            return None
        return pd.Timestamp(t).to_period("Q").end_time.date()

    def _parse_first_evidence(row: pd.Series) -> Dict[str, Any]:
        for key in ["evidence_json", "source_evidence_json", "evidence"]:
            raw = row.get(key)
            if isinstance(raw, str) and raw.strip():
                try:
                    parsed = json.loads(raw)
                    if isinstance(parsed, list) and parsed:
                        first = parsed[0]
                        if isinstance(first, dict):
                            return first
                    if isinstance(parsed, dict):
                        return parsed
                except Exception:
                    continue
        return {}

    def _analysis_sheet_layout_support_runtime() -> Dict[str, Any]:
        return {
            "_get_analysis_sheet_style_bundle": _get_analysis_sheet_style_bundle,
            "copy": copy,
            "Font": Font,
            "Alignment": Alignment,
            "Border": Border,
            "header_size": header_size,
        }

    def _analysis_sheet_layout_support() -> AnalysisSheetLayoutSupport:
        return AnalysisSheetLayoutSupport(
            AnalysisSheetLayoutSupportDeps(runtime=_analysis_sheet_layout_support_runtime())
        )

    def _write_analysis_sheet_title_and_metadata(
        ws: Any,
        title: str,
        metadata_text: str,
        *,
        max_col: int,
        title_row: int = 1,
        metadata_row: int = 2,
    ) -> int:
        return _analysis_sheet_layout_support().write_analysis_sheet_title_and_metadata(
            ws,
            title,
            metadata_text,
            max_col=max_col,
            title_row=title_row,
            metadata_row=metadata_row,
        )

    def _render_stacked_quarter_blocks(
        ws: Any,
        quarters: List[date],
        rows_by_quarter: Dict[date, List[Dict[str, Any]]],
        max_col: int,
        block_title_fn: Any,
        row_writer: Any,
        block_header_writer: Optional[Any] = None,
        start_row: int = 2,
        blank_row_between: bool = True,
    ) -> int:
        return _analysis_sheet_layout_support().render_stacked_quarter_blocks(
            ws,
            quarters,
            rows_by_quarter,
            max_col,
            block_title_fn,
            row_writer,
            block_header_writer=block_header_writer,
            start_row=start_row,
            blank_row_between=blank_row_between,
        )

    _quarter_notes_ui_selection_outer_scope = dict(locals())

    def _quarter_notes_context_adapter_deps() -> QuarterNotesContextAdapterDeps:
        return QuarterNotesContextAdapterDeps(
            runtime={
                "QuarterNotesUiOrchestratorDeps": QuarterNotesUiOrchestratorDeps,
                "write_quarter_notes_ui_sheet": write_quarter_notes_ui_sheet,
                "wb": wb,
                "ticker": ticker,
                "company_profile": company_profile,
                "is_pbi_profile": is_pbi_profile,
                "is_gpre_profile": is_gpre_profile,
                "is_anf_profile": is_anf_profile,
                "quarter_notes": quarter_notes,
                "hist": hist,
                "promises": promises,
                "cache_root": cache_root,
                "inputs": inputs,
                "ui_state": ui_state,
                "ui_info_rows": ui_info_rows,
                "ctx_ref": ctx_ref,
                "quarter_notes_runtime": quarter_notes_runtime,
                "context_globals": globals(),
                "quarter_notes_ui_selection_outer_scope": _quarter_notes_ui_selection_outer_scope,
                "write_analysis_sheet_title_and_metadata": _write_analysis_sheet_title_and_metadata,
                "get_analysis_sheet_style_bundle": _get_analysis_sheet_style_bundle,
                "quarter_notes_view": _quarter_notes_view,
                "resolve_col": _resolve_col,
                "normalize_text": glx_normalize_text,
                "split_sentences": glx_split_sentences,
                "dedup_text_key": glx_dedup_text_key,
                "extract_numeric_patterns": glx_extract_numeric_patterns,
                "normalize_period": glx_normalize_period,
                "compact_snippet": qn_compact_snippet,
                "quarter_label_short": _quarter_label_short,
                "ensure_terminal_period": _ensure_terminal_period,
                "collapse_repeated_leading_ngram": _collapse_repeated_leading_ngram_local,
                "dedupe_canonical_text_parts": _dedupe_canonical_text_parts_local,
                "quarter_note_runtime_qd_token": _quarter_note_runtime_qd_token,
                "quarter_note_runtime_signature": _quarter_note_runtime_signature,
                "quarter_note_runtime_cache_key": _quarter_note_runtime_cache_key,
                "shared_build_evidence_event": shared_build_evidence_event,
                "audit_view": _audit_view,
                "submission_recent_rows": _submission_recent_rows,
                "submission_recent_row_quarter": _submission_recent_row_quarter,
                "sec_docs_for_accession": _sec_docs_for_accession,
                "resolve_cached_doc_path": _resolve_cached_doc_path,
                "path_cache_key": _path_cache_key,
                "read_cached_doc_text": _read_cached_doc_text,
                "parse_date": parse_date,
                "anf_visible_quarter_note_summaries": _anf_visible_quarter_note_summaries,
                "anf_clean_visible_ui_text": _anf_clean_visible_ui_text,
                "anf_polish_quarter_note_visible_fields": _anf_polish_quarter_note_visible_fields,
                "record_writer_substage": _record_writer_substage,
                "timed_writer_substage": _timed_writer_substage,
                "record_writer_elapsed": _record_writer_elapsed,
                "quarter_narrative_recent_periods_from_frame": _quarter_narrative_recent_periods_from_frame,
                "quarter_narrative_records_for_context": _quarter_narrative_records_for_context,
                "write_quarter_narrative_data_sheet": _write_quarter_narrative_data_sheet,
                "write_quarter_notes_ui_narrative_sheet": _write_quarter_notes_ui_narrative_sheet,
            }
        )

    def _write_quarter_notes_ui_v2(
        rank_cutoff: int = 8, severity_cutoff: float = 50.0, max_rows_per_category: int = 10, quarters_shown: int = 12
    ) -> List[Dict[str, Any]]:
        return _write_quarter_notes_ui_v2_impl(
            _quarter_notes_context_adapter_deps(),
            rank_cutoff=rank_cutoff,
            severity_cutoff=severity_cutoff,
            max_rows_per_category=max_rows_per_category,
            quarters_shown=quarters_shown,
        )

    def _write_promise_tracker_ui_v2(render_visible: bool = True) -> List[Dict[str, Any]]:
        deps = PromiseTrackerWriterDeps(
            wb=wb,
            promises=promises,
            slides_guidance=slides_guidance,
            promise_evidence_df=promise_evidence_df,
            ui_state=ui_state,
            ui_info_rows=ui_info_rows,
            company_profile=company_profile,
            ticker=ticker,
            is_pbi_profile=is_pbi_profile,
            is_gpre_profile=is_gpre_profile,
            header_size=header_size,
            apply_hyperlink_look=_apply_hyperlink_look,
            candidate_quality_key=_candidate_quality_key,
            classify_pbi_metric_label=_classify_pbi_metric_label,
            clean_target_bonus=_clean_target_bonus,
            derive_split_target_meta=_derive_split_target_meta,
            extract_45z_monetization_target_display=_extract_45z_monetization_target_display,
            extract_money_targets_for_display=_extract_money_targets_for_display,
            extract_pbi_guidance_targets_multi=_extract_pbi_guidance_targets_multi,
            extract_pbi_target_display=_extract_pbi_target_display,
            fmt_short_money_value_local=_fmt_short_money_value_local,
            gpre_bad_visible_promise_reason=_gpre_bad_visible_promise_reason,
            gpre_clean_visible_promise_metric=_gpre_clean_visible_promise_metric,
            is_45z_crush_margin_support_only=_is_45z_crush_margin_support_only,
            is_pbi_clean_sentence=_is_pbi_clean_sentence,
            is_preferred_narrative_source=_is_preferred_narrative_source,
            load_profile_slide_signals=_load_profile_slide_signals,
            looks_pbi_fragment_text=_looks_pbi_fragment_text,
            management_theme_key=_management_theme_key,
            pbi_promise_theme_re=_pbi_promise_theme_re,
            pbi_structured_guidance_items_for_qd=_pbi_structured_guidance_items_for_qd,
            pbi_structured_strategy_items_for_qd=_pbi_structured_strategy_items_for_qd,
            pbi_target_display_ok=_pbi_target_display_ok,
            profile_slide_metric=_profile_slide_metric,
            promises_view=_promises_view,
            resolve_col=_resolve_col,
            set_cell_comment=_set_cell_comment_local,
            slide_signal_noise=_slide_signal_noise,
            source_rank=_source_rank,
            split_target_group_key=_split_target_group_key,
            split_target_identity_key=_split_target_identity_key,
            split_target_metric_display=_split_target_metric_display,
            split_target_scope_token=_split_target_scope_token,
            strong_45z_2026_target_display=_strong_45z_2026_target_display,
            text_fragment_penalty=_text_fragment_penalty,
        )
        return write_promise_tracker_ui_sheet(deps, render_visible=render_visible)

    def _ensure_promise_progress_ui_bundle(
        quarter_hint: Optional[Tuple[date, ...]] = None,
    ) -> Dict[str, Any]:
        nonlocal promise_progress_ui_bundle_cache
        deps = PromiseProgressUiBundleDeps(
            promise_progress=promise_progress,
            promise_evidence_df=promise_evidence_df,
            hist=hist,
            adj_metrics=adj_metrics,
            is_pbi_profile=is_pbi_profile,
            resolve_col=_resolve_col,
            hist_view=_hist_view,
            adj_metrics_view=_adj_metrics_view,
            classify_pbi_metric_label=_classify_pbi_metric_label,
            extract_pbi_target_display=_extract_pbi_target_display,
            extract_45z_monetization_target_display=_extract_45z_monetization_target_display,
            strong_45z_2026_target_display=_strong_45z_2026_target_display,
            extract_money_targets_for_display=_extract_money_targets_for_display,
            fmt_short_money_value_local=_fmt_short_money_value_local,
        )
        bundle = build_promise_progress_ui_bundle(
            deps,
            quarter_hint=quarter_hint,
            cached_bundle=promise_progress_ui_bundle_cache,
        )
        promise_progress_ui_bundle_cache = bundle
        if ctx_ref is not None:
            ctx_ref.derived.promise_progress_ui_bundle = bundle
        return bundle

    def _write_promise_progress_ui_v2() -> List[Dict[str, Any]]:
        return promise_progress_write_promise_progress_ui_v2(
            PromiseProgressOrchestratorDeps(
                wb=wb,
                ticker=ticker,
                is_anf_profile=is_anf_profile,
                is_pbi_profile=is_pbi_profile,
                is_gpre_profile=is_gpre_profile,
                promise_progress=promise_progress,
                promises=promises,
                ui_state=ui_state if isinstance(ui_state, dict) else {},
                ui_info_rows=ui_info_rows,
                hist=hist,
                adj_metrics=adj_metrics,
                slides_guidance=slides_guidance,
                material_roots=material_roots,
                ticker_roots=ticker_roots,
                pdf_text_cache_root=pdf_text_cache_root,
                rebuild_doc_text_cache=rebuild_doc_text_cache,
                quiet_pdf_warnings=quiet_pdf_warnings,
                quarter_notes=quarter_notes,
                promise_visible_max_col=PROMISE_VISIBLE_MAX_COL,
                promise_timeline_headers=PROMISE_TIMELINE_HEADERS,
                anf_build_promise_progress_sections=_anf_build_promise_progress_sections,
                anf_clean_visible_ui_text=_anf_clean_visible_ui_text,
                apply_hyperlink_look=_apply_hyperlink_look,
                candidate_quality_key=_candidate_quality_key,
                classify_pbi_metric_label=_classify_pbi_metric_label,
                clean_target_bonus=_clean_target_bonus,
                coerce_amount_with_unit_local=_coerce_amount_with_unit_local,
                derive_split_target_meta=_derive_split_target_meta,
                ensure_promise_progress_ui_bundle=_ensure_promise_progress_ui_bundle,
                ensure_terminal_period=_ensure_terminal_period,
                estimate_wrapped_line_count=_estimate_wrapped_line_count,
                estimate_wrapped_row_height=_estimate_wrapped_row_height,
                excel_safe_text_local=_excel_safe_text_local,
                extract_45z_monetization_target_display=_extract_45z_monetization_target_display,
                extract_45z_realized_progress_text=_extract_45z_realized_progress_text,
                extract_money_targets_for_display=_extract_money_targets_for_display,
                extract_pbi_target_display=_extract_pbi_target_display,
                fmt_short_money_value_local=_fmt_short_money_value_local,
                get_analysis_sheet_style_bundle=_get_analysis_sheet_style_bundle,
                gpre_bad_visible_promise_reason=_gpre_bad_visible_promise_reason,
                gpre_clean_visible_promise_metric=_gpre_clean_visible_promise_metric,
                infer_target_period=_infer_target_period,
                infer_target_structure=_infer_target_structure,
                is_45z_crush_margin_support_only=_is_45z_crush_margin_support_only,
                is_pbi_clean_sentence=_is_pbi_clean_sentence,
                is_preferred_narrative_source=_is_preferred_narrative_source,
                load_profile_slide_signals=_load_profile_slide_signals,
                local_slide_45z_realized_text=_local_slide_45z_realized_text,
                looks_pbi_fragment_text=_looks_pbi_fragment_text,
                lookup_pbi_structured_guidance_target=_lookup_pbi_structured_guidance_target,
                lookup_pbi_structured_progress_hint=_lookup_pbi_structured_progress_hint,
                management_credibility_scorecard_rows=_management_credibility_scorecard_rows,
                management_theme_key=_management_theme_key,
                nearest_amount_for_pattern=_nearest_amount_for_pattern,
                parse_quarter_from_filename=_parse_quarter_from_filename,
                parse_quarter_from_follow_text=_parse_quarter_from_follow_text,
                pbi_guidance_period_label_from_text=_pbi_guidance_period_label_from_text,
                pbi_promise_theme_re=_pbi_promise_theme_re,
                pbi_repair_guidance_period_meta=_pbi_repair_guidance_period_meta,
                pbi_structured_strategy_items_for_qd=_pbi_structured_strategy_items_for_qd,
                pbi_target_display_ok=_pbi_target_display_ok,
                quarter_label_short=_quarter_label_short,
                quarter_notes_view=_quarter_notes_view,
                read_cached_doc_raw=_read_cached_doc_raw,
                record_writer_substage=_record_writer_substage,
                render_stacked_quarter_blocks=_render_stacked_quarter_blocks,
                resolve_col=_resolve_col,
                rewrite_shared_promise_progress_ui_from_blocks=_rewrite_shared_promise_progress_ui_from_blocks,
                safe_cell=_safe_cell,
                set_cell_comment_local=_set_cell_comment_local,
                slide_signal_noise=_slide_signal_noise,
                slide_text_paths=_slide_text_paths,
                source_rank=_source_rank,
                split_target_family_key=_split_target_family_key,
                split_target_identity_key=_split_target_identity_key,
                split_target_metric_display=_split_target_metric_display,
                split_target_qend=_split_target_qend,
                split_target_scope_is_broad=_split_target_scope_is_broad,
                split_target_scope_token=_split_target_scope_token,
                strong_45z_2026_target_display=_strong_45z_2026_target_display,
                target_period_is_closed=_target_period_is_closed,
                text_fragment_penalty=_text_fragment_penalty,
                timed_writer_substage=_timed_writer_substage,
                write_analysis_sheet_title_and_metadata=_write_analysis_sheet_title_and_metadata,
            )
        )

    financial_report_support: Optional[FinancialReportSupport] = None

    def _financial_report_support_runtime() -> Dict[str, Any]:
        return {
            "wb": wb,
            "hist": hist,
            "audit": audit,
            "needs_review": needs_review,
            "company_profile": company_profile,
            "bank_metrics_enabled": bank_metrics_enabled,
            "font_size": font_size,
            "header_size": header_size,
            "pd": pd,
            "date": date,
            "datetime": datetime,
            "Font": Font,
            "Alignment": Alignment,
            "Table": Table,
            "TableStyleInfo": TableStyleInfo,
            "get_column_letter": get_column_letter,
            "classify_duration": classify_duration,
            "_hist_view": _hist_view,
            "_audit_view": _audit_view,
            "_source_class": _source_class,
            "_source_method": _source_method,
            "_source_qa": _source_qa,
            "_source_tier": _source_tier,
            "_source_label": _source_label,
            "_safe_cell": _safe_cell,
            "_autowidth": _autowidth,
            "_updated_font": _updated_font,
            "strictness": strictness,
        }

    def _get_financial_report_support() -> FinancialReportSupport:
        nonlocal financial_report_support
        if financial_report_support is None:
            financial_report_support = FinancialReportSupport(
                FinancialReportSupportDeps(runtime=_financial_report_support_runtime())
            )
        return financial_report_support

    def _period_type(row: pd.Series) -> str:
        return _get_financial_report_support().period_type(row)

    def _build_facts_long() -> pd.DataFrame:
        return _get_financial_report_support().build_facts_long()

    def _build_lineitem_map() -> pd.DataFrame:
        return _get_financial_report_support().build_lineitem_map()

    def _build_period_index(max_periods: int = 12) -> pd.DataFrame:
        return _get_financial_report_support().build_period_index(max_periods)

    def _build_report(statement: str, scale: float = 1e6) -> pd.DataFrame:
        return _get_financial_report_support().build_report(statement, scale=scale)

    def _write_report_sheet(name: str, df: pd.DataFrame, scale_label: str) -> None:
        return _get_financial_report_support().write_report_sheet(name, df, scale_label)

    def _write_summary_sheet(df: pd.DataFrame) -> None:
        return write_summary_sheet(
            SummarySheetRenderDeps(
                wb=wb,
                font_size=font_size,
                header_size=header_size,
                set_cell_comment=_set_cell_comment_local,
                normalize_text=glx_normalize_text,
                estimate_wrapped_line_count=_estimate_wrapped_line_count,
                estimate_wrapped_row_height=_estimate_wrapped_row_height,
            ),
            df,
        )
    _gpre_commercial_setup_support_cache: Optional[GpreCommercialSetupSupport] = None

    def _get_gpre_commercial_setup_support() -> GpreCommercialSetupSupport:
        nonlocal _gpre_commercial_setup_support_cache
        if _gpre_commercial_setup_support_cache is None:
            _gpre_commercial_setup_support_cache = GpreCommercialSetupSupport(
                GpreCommercialSetupDeps(
                    is_gpre_profile=is_gpre_profile,
                    ctx_ref=ctx_ref,
                    cache_dir=cache_dir,
                    load_operating_driver_source_records_by_quarter=_load_operating_driver_source_records_by_quarter,
                    normalize_text=glx_normalize_text,
                    split_sentences=glx_split_sentences,
                    compact_snippet=qn_compact_snippet,
                    ensure_terminal_period=_ensure_terminal_period,
                    data_root_from_sec_cache_path=data_root_from_sec_cache_path,
                )
            )
        return _gpre_commercial_setup_support_cache

    def _read_local_doc_text_shared(path_in: Any) -> str:
        return _get_gpre_commercial_setup_support().read_local_doc_text(path_in)

    def _gpre_local_bofa_conference_path_shared() -> Path:
        return _get_gpre_commercial_setup_support().local_bofa_conference_path()

    def _gpre_local_bofa_conference_text_shared() -> str:
        return _get_gpre_commercial_setup_support().local_bofa_conference_text()

    def _gpre_local_stephens_conference_path_shared() -> Path:
        return _get_gpre_commercial_setup_support().local_stephens_conference_path()

    def _gpre_local_stephens_conference_raw_path_shared() -> Path:
        return _get_gpre_commercial_setup_support().local_stephens_conference_raw_path()

    def _gpre_local_stephens_conference_text_shared() -> str:
        return _get_gpre_commercial_setup_support().local_stephens_conference_text()

    def _gpre_local_stephens_conference_raw_text_shared() -> str:
        return _get_gpre_commercial_setup_support().local_stephens_conference_raw_text()

    def _gpre_local_bmo_conference_path_shared() -> Path:
        return _get_gpre_commercial_setup_support().local_bmo_conference_path()

    def _gpre_local_bmo_conference_text_shared() -> str:
        return _get_gpre_commercial_setup_support().local_bmo_conference_text()

    def _gpre_local_bofa_conference_excerpt_shared(terms: Tuple[str, ...], max_len: int = 280) -> str:
        return _get_gpre_commercial_setup_support().local_bofa_conference_excerpt(terms, max_len)

    def _gpre_commercial_setup_records_shared() -> List[Dict[str, Any]]:
        return _get_gpre_commercial_setup_support().records()

    def _write_valuation_sheet() -> None:
        valuation_orchestrator_helpers = {
            "Alignment": Alignment,
            "Border": Border,
            "CellIsRule": CellIsRule,
            "DefinedName": DefinedName,
            "Font": Font,
            "Path": Path,
            "PatternFill": PatternFill,
            "Side": Side,
            "_anf_buyback_execution_is_year_or_ttm": _anf_buyback_execution_is_year_or_ttm,
            "_anf_format_year_ttm_buyback_summary": _anf_format_year_ttm_buyback_summary,
            "_anf_is_missing_value": _anf_is_missing_value,
            "_anf_normalize_ytd_buyback_cash_map_for_valuation": _anf_normalize_ytd_buyback_cash_map_for_valuation,
            "_anf_prior_year_quarter": _anf_prior_year_quarter,
            "_anf_value_delta_map_for_fiscal_periods": _anf_value_delta_map_for_fiscal_periods,
            "_anf_visible_quarter_label": _anf_visible_quarter_label,
            "_anf_yoy_map_for_fiscal_periods": _anf_yoy_map_for_fiscal_periods,
            "_audit_view": _audit_view,
            "_build_hidden_value_flags_fallback": _build_hidden_value_flags_fallback,
            "_build_operating_driver_rows": _build_operating_driver_rows,
            "_collapse_repeated_leading_ngram_local": _collapse_repeated_leading_ngram_local,
            "_dedupe_canonical_text_parts_local": _dedupe_canonical_text_parts_local,
            "_ensure_terminal_period": _ensure_terminal_period,
            "_ensure_valuation_precompute_bundle": _ensure_valuation_precompute_bundle,
            "_ensure_valuation_render_bundle": _ensure_valuation_render_bundle,
            "_estimate_wrapped_row_height": _estimate_wrapped_row_height,
            "_extract_45z_monetization_target_display": _extract_45z_monetization_target_display,
            "_extract_latest_buyback_remaining_from_sec": _extract_latest_buyback_remaining_from_sec,
            "_extract_money_targets_for_display": _extract_money_targets_for_display,
            "_extract_pbi_target_display": _extract_pbi_target_display,
            "_extract_valuation_filing_doc_text": _extract_valuation_filing_doc_text,
            "_first_existing_material_dir": _first_existing_material_dir,
            "_fmt_short_money_value_local": _fmt_short_money_value_local,
            "_gpre_commercial_setup_records_shared": _gpre_commercial_setup_records_shared,
            "_gpre_local_bmo_conference_path_shared": _gpre_local_bmo_conference_path_shared,
            "_gpre_local_bmo_conference_text_shared": _gpre_local_bmo_conference_text_shared,
            "_gpre_local_bofa_conference_path_shared": _gpre_local_bofa_conference_path_shared,
            "_gpre_local_bofa_conference_text_shared": _gpre_local_bofa_conference_text_shared,
            "_gpre_local_stephens_conference_path_shared": _gpre_local_stephens_conference_path_shared,
            "_gpre_local_stephens_conference_raw_path_shared": _gpre_local_stephens_conference_raw_path_shared,
            "_gpre_local_stephens_conference_raw_text_shared": _gpre_local_stephens_conference_raw_text_shared,
            "_gpre_local_stephens_conference_text_shared": _gpre_local_stephens_conference_text_shared,
            "_gpre_normalize_metric_label": _gpre_normalize_metric_label,
            "_htmlish_to_text": _htmlish_to_text,
            "_load_profile_slide_signals": _load_profile_slide_signals,
            "_operating_driver_financial_statement_files": _operating_driver_financial_statement_files,
            "_parse_quarter_from_filename": _parse_quarter_from_filename,
            "_parse_quarter_from_follow_text": _parse_quarter_from_follow_text,
            "_pbi_guidance_period_label_from_text": _pbi_guidance_period_label_from_text,
            "_pbi_repair_guidance_period_meta": _pbi_repair_guidance_period_meta,
            "_pbi_structured_strategy_items_for_qd": _pbi_structured_strategy_items_for_qd,
            "_period_label_to_norm": _period_label_to_norm,
            "_prev_quarter_end_from_qend": _prev_quarter_end_from_qend,
            "_profile_slide_signals_for_quarter": _profile_slide_signals_for_quarter,
            "_promises_view": _promises_view,
            "_quarter_notes_view": _quarter_notes_view,
            "_read_cached_doc_text": _read_cached_doc_text,
            "_read_local_doc_text_shared": _read_local_doc_text_shared,
            "_read_operating_driver_text": _read_operating_driver_text,
            "_resolve_cached_doc_path": _resolve_cached_doc_path,
            "_resolve_col": _resolve_col,
            "_resolve_thesis_fy_base": _resolve_thesis_fy_base,
            "_safe_text_value": _safe_text_value,
            "_sec_cache_docs_for_token_local": _sec_cache_docs_for_token_local,
            "_sec_docs_for_accession": _sec_docs_for_accession,
            "_slide_signal_noise": _slide_signal_noise,
            "_source_backed_debt_tranches_from_slides": _source_backed_debt_tranches_from_slides,
            "_submission_recent_row_quarter": _submission_recent_row_quarter,
            "_submission_recent_rows": _submission_recent_rows,
            "_updated_font": _updated_font,
            "annual_segment_alias_patterns": annual_segment_alias_patterns,
            "build_hidden_value_flags": build_hidden_value_flags,
            "build_valuation_history_source_maps": build_valuation_history_source_maps,
            "copy": copy,
            "display_m_source_map": display_m_source_map,
            "ew_latest_segment_financials_workbook": ew_latest_segment_financials_workbook,
            "ew_parse_quarterly_segment_data_from_workbook": ew_parse_quarterly_segment_data_from_workbook,
            "font_size": font_size,
            "get_column_letter": get_column_letter,
            "glx_normalize_text": glx_normalize_text,
            "header_size": header_size,
            "history_margin_source_map": history_margin_source_map,
            "history_numeric_source_map": history_numeric_source_map,
            "infer_quarter_end_from_text": infer_quarter_end_from_text,
            "normalize_capex_for_valuation": normalize_capex_for_valuation,
            "pd": pd,
            "qn_compact_snippet": qn_compact_snippet,
            "quarter_key_union": quarter_key_union,
            "re": re,
            "source_infer_q_from_name": source_infer_q_from_name,
            "strip_html": strip_html,
            "ttm_map": ttm_map,
            "ttm_sparse_cashflow_map": ttm_sparse_cashflow_map,
            "valuation_hidden_comparison_metric": valuation_hidden_comparison_metric,
        }
        return write_valuation_sheet(
            ValuationOrchestratorDeps(
                wb=wb,
                ticker=ticker,
                company_profile=company_profile,
                is_pbi_profile=is_pbi_profile,
                is_gpre_profile=is_gpre_profile,
                is_anf_profile=is_anf_profile,
                price=price,
                excel_mode=excel_mode,
                hist=hist,
                quarter_notes=quarter_notes,
                promises=promises,
                audit=audit,
                promise_progress=promise_progress,
                slides_guidance=slides_guidance,
                slides_debt=slides_debt,
                valuation_grid_df=valuation_grid_df,
                adj_metrics=adj_metrics,
                adj_metrics_relaxed=adj_metrics_relaxed,
                leverage_df=leverage_df,
                manifest_df=manifest_df,
                flags_df=flags_df,
                flags_audit_df=flags_audit_df,
                signals_base_df=signals_base_df,
                debt_tranches=debt_tranches,
                debt_tranches_latest=debt_tranches_latest,
                debt_credit_notes=debt_credit_notes,
                company_overview=company_overview,
                cache_root=cache_root,
                cache_dir=cache_dir,
                material_roots=material_roots,
                ctx_ref=ctx_ref,
                ui_state=ui_state,
                context_globals=globals(),
                get_valuation_style_bundle=_get_valuation_style_bundle,
                set_cell_comment=_set_cell_comment_local,
                timed_writer_substage=_timed_writer_substage,
                record_writer_substage=_record_writer_substage,
                record_writer_elapsed=_record_writer_elapsed,
                context_helpers=valuation_orchestrator_helpers,
            )
        )

    def _build_summary() -> pd.DataFrame:
        return build_summary_dataframe(
            SummaryBuilderDeps(
                hist=hist,
                leverage_df=leverage_df,
                needs_review=needs_review,
                company_overview=company_overview,
                price=price,
                ctx_ref=ctx_ref,
                hist_view=_hist_view,
                audit_view=_audit_view,
            )
        )

    facts_long = pd.DataFrame()
    lineitem_map = pd.DataFrame()
    period_index = pd.DataFrame()
    report_is = pd.DataFrame()
    report_bs = pd.DataFrame()
    report_cf = pd.DataFrame()

    def _normalize_leverage_text(text_in: Any) -> str:
        return source_normalize_leverage_text(text_in)

    def _infer_q_from_name(name: str) -> Optional[dt.date]:
        return source_infer_q_from_name(name)

    def _normalize_leverage_quarter(qv: Any) -> Optional[pd.Timestamp]:
        return source_normalize_leverage_quarter(qv)

    def _hist_quarter_whitelist() -> Set[pd.Timestamp]:
        return source_hist_quarter_whitelist(hist)

    def _read_leverage_material_text(path_in: Path) -> str:
        return _read_cached_doc_text(path_in, normalize=True)

    def _looks_like_leverage_text(text_in: Any) -> bool:
        return source_looks_like_leverage_text(text_in)

    def _load_leverage_local_material_index() -> List[Dict[str, Any]]:
        nonlocal leverage_local_material_index_cache
        if leverage_local_material_index_cache is not None:
            if ctx_ref is not None:
                ctx_ref.derived.valuation_local_material_index = list(leverage_local_material_index_cache)
            return leverage_local_material_index_cache
        leverage_local_material_index_cache = source_build_leverage_local_material_index(
            hist=hist,
            material_roots=tuple(material_roots),
            ticker=ticker,
            ticker_roots=tuple(ticker_roots),
            read_cached_doc_text_fn=_read_cached_doc_text,
            infer_cached_doc_quarter_fn=_infer_cached_doc_quarter,
        )
        if ctx_ref is not None:
            ctx_ref.derived.valuation_local_material_index = list(leverage_local_material_index_cache)
        return leverage_local_material_index_cache

    def _load_leverage_audit_doc_index() -> List[Dict[str, Any]]:
        nonlocal leverage_audit_doc_index_cache
        if leverage_audit_doc_index_cache is not None:
            if ctx_ref is not None:
                ctx_ref.derived.valuation_audit_doc_index = list(leverage_audit_doc_index_cache)
            return leverage_audit_doc_index_cache
        rows: List[Dict[str, Any]] = []
        if audit is None or audit.empty:
            leverage_audit_doc_index_cache = rows
            if ctx_ref is not None:
                ctx_ref.derived.valuation_audit_doc_index = list(rows)
            return rows
        leverage_audit_doc_index_cache = source_build_leverage_audit_doc_index(
            audit=audit,
            cache_root=cache_root,
            resolve_col=_resolve_col,
            accession_doc_lookup=_sec_docs_for_accession,
            read_leverage_material_text_fn=_read_leverage_material_text,
        )
        if ctx_ref is not None:
            ctx_ref.derived.valuation_audit_doc_index = list(leverage_audit_doc_index_cache)
        return leverage_audit_doc_index_cache

    def _extract_adj_net_leverage_text_map() -> Dict[pd.Timestamp, float]:
        nonlocal adj_net_leverage_text_map_cache
        if adj_net_leverage_text_map_cache is not None:
            if ctx_ref is not None:
                ctx_ref.derived.valuation_net_leverage_text_map = dict(adj_net_leverage_text_map_cache)
            return dict(adj_net_leverage_text_map_cache)
        out_map = source_extract_adj_net_leverage_text_map(
            promises=promises,
            quarter_notes=quarter_notes,
            slides_guidance=slides_guidance,
            ocr_log=ocr_log,
            hist=hist,
            resolve_col=_resolve_col,
            load_local_material_index_fn=_load_leverage_local_material_index,
            load_audit_doc_index_fn=_load_leverage_audit_doc_index,
            timed_substage=_timed_writer_substage,
        )
        adj_net_leverage_text_map_cache = dict(out_map)
        if ctx_ref is not None:
            ctx_ref.derived.valuation_net_leverage_text_map = dict(out_map)
            if leverage_local_material_index_cache is not None:
                ctx_ref.derived.valuation_local_material_index = list(leverage_local_material_index_cache)
            if leverage_audit_doc_index_cache is not None:
                ctx_ref.derived.valuation_audit_doc_index = list(leverage_audit_doc_index_cache)
        return out_map


    # Non-GAAP bridge (strict + relaxed)
    def _build_ng_bridge(adj_df: pd.DataFrame, breakdown_df: pd.DataFrame) -> pd.DataFrame:
        if hist is None or hist.empty:
            return pd.DataFrame()
        if adj_df is None or adj_df.empty:
            return pd.DataFrame()
        h = hist.copy()
        h["quarter"] = pd.to_datetime(h["quarter"], errors="coerce")
        adj = adj_df.copy()
        if "quarter" not in adj.columns:
            return pd.DataFrame()
        adj["quarter"] = pd.to_datetime(adj["quarter"], errors="coerce")
        rows = []
        for q in sorted(adj["quarter"].dropna().unique()):
            hq = h[h["quarter"] == q]
            if hq.empty:
                continue
            gaap_op = pd.to_numeric(hq["op_income"], errors="coerce").iloc[0] if "op_income" in hq.columns else None
            gaap_ebitda = pd.to_numeric(hq["ebitda"], errors="coerce").iloc[0] if "ebitda" in hq.columns else None
            sub = adj[adj["quarter"] == q]
            adj_ebit = pd.to_numeric(sub.get("adj_ebit"), errors="coerce").iloc[0] if "adj_ebit" in sub else None
            adj_ebitda = pd.to_numeric(sub.get("adj_ebitda"), errors="coerce").iloc[0] if "adj_ebitda" in sub else None
            adj_sum = None
            if breakdown_df is not None and not breakdown_df.empty:
                bd = breakdown_df.copy()
                if "quarter" in bd.columns:
                    bd["quarter"] = pd.to_datetime(bd["quarter"], errors="coerce")
                    bdq = bd[bd["quarter"] == q]
                    if not bdq.empty and "value" in bdq.columns:
                        adj_sum = pd.to_numeric(bdq["value"], errors="coerce").sum()
            rows.append({
                "quarter": q.date(),
                "gaap_op_income": gaap_op,
                "adj_ebit": adj_ebit,
                "adjustments_sum": adj_sum,
                "gaap_ebitda": gaap_ebitda,
                "adj_ebitda": adj_ebitda,
            })
        return pd.DataFrame(rows)

    leverage_df = pd.DataFrame()
    valuation_summary_df = pd.DataFrame()
    valuation_grid_df = pd.DataFrame()
    summary_df = pd.DataFrame()
    signals_base_df = pd.DataFrame()
    flags_df = pd.DataFrame()
    flags_audit_df = pd.DataFrame()
    flags_recompute_df = pd.DataFrame()
    ng_bridge = pd.DataFrame()
    ng_bridge_relaxed = pd.DataFrame()

    data_is_rules_df = pd.DataFrame()
    if is_rules:
        rows = []
        for k, v in is_rules.items():
            if isinstance(v, list):
                rows.append({"rule": k, "value": "; ".join(str(x) for x in v)})
            else:
                rows.append({"rule": k, "value": v})
        data_is_rules_df = pd.DataFrame(rows)

    def _evidence_source_support_runtime() -> Dict[str, Any]:
        return {
            "pd": pd,
            "json": json,
            "hashlib": hashlib,
            "quarter_notes": quarter_notes,
            "promise_progress": promise_progress,
            "promises": promises,
            "_quarter_notes_view": _quarter_notes_view,
            "_promises_view": _promises_view,
            "_resolve_col": _resolve_col,
            "_qend_date": _qend_date,
            "_parse_first_evidence": _parse_first_evidence,
        }

    def _evidence_source_support() -> EvidenceSourceSupport:
        return EvidenceSourceSupport(EvidenceSourceSupportDeps(runtime=_evidence_source_support_runtime()))

    def _build_qn_evidence_src() -> pd.DataFrame:
        return _evidence_source_support().build_qn_evidence_src()

    def _build_promise_evidence_src() -> pd.DataFrame:
        return _evidence_source_support().build_promise_evidence_src()

    quarter_notes_evidence_df = pd.DataFrame()
    promise_evidence_df = pd.DataFrame()
    ui_state = {"quarters": [], "promise_rows": pd.DataFrame()}

    _latest_quarter_qa_support_cache: Optional[LatestQuarterQASupport] = None

    def _get_latest_quarter_qa_support() -> LatestQuarterQASupport:
        nonlocal _latest_quarter_qa_support_cache
        if _latest_quarter_qa_support_cache is None:
            _latest_quarter_qa_support_cache = LatestQuarterQASupport(
                LatestQuarterQADeps(
                    ticker=ticker,
                    company_profile=company_profile,
                    is_pbi_profile=is_pbi_profile,
                    is_gpre_profile=is_gpre_profile,
                    is_anf_profile=is_anf_profile,
                    hist=hist,
                    leverage_df=leverage_df,
                    adj_metrics=adj_metrics,
                    audit=audit,
                    slides_guidance=slides_guidance,
                    slides_segments=slides_segments,
                    debt_tranches_latest=debt_tranches_latest,
                    debt_buckets=debt_buckets,
                    debt_profile=debt_profile,
                    debt_recon=debt_recon,
                    revolver_history=revolver_history,
                    non_gaap_files=non_gaap_files,
                    cache_root=cache_root,
                    material_roots=material_roots,
                    ticker_roots=ticker_roots,
                    document_cache=document_cache,
                    ui_state=ui_state,
                    ctx_ref=ctx_ref,
                    context_helpers={
                        "_hist_view": _hist_view,
                        "_audit_view": _audit_view,
                        "_adj_metrics_view": _adj_metrics_view,
                        "_resolve_col": _resolve_col,
                        "_submission_recent_rows": _submission_recent_rows,
                        "_submission_recent_row_quarter": _submission_recent_row_quarter,
                        "_sec_docs_for_accession": _sec_docs_for_accession,
                        "_resolve_cached_doc_path": _resolve_cached_doc_path,
                        "_read_cached_doc_text": _read_cached_doc_text,
                        "_path_belongs_to_ticker": _path_belongs_to_ticker,
                        "_first_existing_material_dir": _first_existing_material_dir,
                        "_ensure_valuation_render_bundle": _ensure_valuation_render_bundle,
                        "_ensure_valuation_precompute_bundle": _ensure_valuation_precompute_bundle,
                        "_timed_writer_substage": _timed_writer_substage,
                        "_parse_quarter_from_filename": _parse_quarter_from_filename,
                        "_parse_quarter_from_follow_text": _parse_quarter_from_follow_text,
                        "_pbi_reported_fcf_payload_for_qd": _pbi_reported_fcf_payload_for_qd,
                        "_anf_financial_schedule_support_doc_for_quarter": _anf_financial_schedule_support_doc_for_quarter,
                        "_slides_guidance_has_explicit_metric": _slides_guidance_has_explicit_metric,
                        "coerce_number": coerce_number,
                        "infer_quarter_end_from_text": infer_quarter_end_from_text,
                        "writer_qa_latest_quarter_support_gap_severity": writer_qa_latest_quarter_support_gap_severity,
                        "ew_parse_quarterly_segment_data_from_workbook": ew_parse_quarterly_segment_data_from_workbook,
                        "annual_segment_alias_patterns": annual_segment_alias_patterns,
                        "quarterly_segment_labels": quarterly_segment_labels,
                    },
                )
            )
        return _latest_quarter_qa_support_cache

    def _latest_quarter_qa_source_bundle(
        qref: pd.Timestamp,
        *,
        include_transcripts: bool = False,
    ) -> List[Dict[str, Any]]:
        return _get_latest_quarter_qa_support().source_bundle(
            qref,
            include_transcripts=include_transcripts,
        )

    def _latest_quarter_sec_text_corpus(qref: pd.Timestamp) -> str:
        return _get_latest_quarter_qa_support().sec_text_corpus(qref)

    def _run_latest_quarter_qa() -> List[Dict[str, Any]]:
        return _get_latest_quarter_qa_support().run()

    def _write_bs_segments_sheet(quarters_shown: int = 8) -> List[Dict[str, Any]]:
        deps = BSSegmentsSheetAdapterDeps(
            runtime={
                "BsSegmentsWriterDeps": BsSegmentsWriterDeps,
                "write_bs_segments_sheet": write_bs_segments_sheet,
                "wb": wb,
                "hist": hist,
                "audit": audit,
                "ticker": ticker,
                "company_profile": company_profile,
                "slides_segments": slides_segments,
                "material_roots": material_roots,
                "ticker_roots": ticker_roots,
                "ui_info_rows": ui_info_rows,
                "font_size": font_size,
                "header_size": header_size,
                "is_pbi_profile": is_pbi_profile,
                "is_gpre_profile": is_gpre_profile,
                "is_anf_profile": is_anf_profile,
                "bank_metrics_enabled": bank_metrics_enabled,
                "enable_quarterly_segment_block": enable_quarterly_segment_block,
                "enable_annual_segment_block": enable_annual_segment_block,
                "quarterly_segment_labels": quarterly_segment_labels,
                "annual_segment_labels": annual_segment_labels,
                "annual_segment_alias_patterns": annual_segment_alias_patterns,
                "ANF_SEGMENT_BRAND_EXPLANATION": ANF_SEGMENT_BRAND_EXPLANATION,
                "_get_valuation_style_bundle": _get_valuation_style_bundle,
                "_hist_view": _hist_view,
                "_resolve_col": _resolve_col,
                "_set_cell_comment_local": _set_cell_comment_local,
                "_shared_load_local_balance_sheet_detail_payloads": _shared_load_local_balance_sheet_detail_payloads,
                "_carry_forward_low_change_series": _carry_forward_low_change_series,
                "_first_existing_material_dir": _first_existing_material_dir,
                "_parse_quarter_from_filename": _parse_quarter_from_filename,
                "_parse_quarter_from_follow_text": _parse_quarter_from_follow_text,
                "_read_operating_driver_text": _read_operating_driver_text,
                "_operating_driver_financial_statement_files": _operating_driver_financial_statement_files,
                "_sec_cache_roots_local": _sec_cache_roots_local,
                "_anf_visible_quarter_label": _anf_visible_quarter_label,
            }
        )
        return BSSegmentsSheetAdapter(deps).write_bs_segments_sheet(quarters_shown=quarters_shown)

    operating_drivers_support: Optional[OperatingDriversSupport] = None

    def _operating_drivers_support_runtime() -> Dict[str, Any]:
        try:
            derivative_oci_bridge_df_runtime = derivative_oci_bridge_df
        except NameError:
            derivative_oci_bridge_df_runtime = pd.DataFrame()
        return {
            **globals(),
            "company_profile": company_profile,
            "hist": hist,
            "quarter_notes": quarter_notes,
            "promises": promises,
            "promise_progress": promise_progress,
            "adj_metrics": adj_metrics,
            "slides_segments": slides_segments,
            "is_gpre_profile": is_gpre_profile,
            "is_anf_profile": is_anf_profile,
            "ctx_ref": ctx_ref,
            "material_roots": material_roots,
            "ticker": ticker,
            "ticker_roots": ticker_roots,
            "cache_dir": cache_dir,
            "ui_info_rows": ui_info_rows,
            "operating_drivers_runtime": operating_drivers_runtime,
            "glx_normalize_text": glx_normalize_text,
            "qn_compact_snippet": qn_compact_snippet,
            "qn_is_complete_signal_text": qn_is_complete_signal_text,
            "_hist_view": _hist_view,
            "_load_profile_slide_signals": _load_profile_slide_signals,
            "_profile_slide_signals_for_quarter": _profile_slide_signals_for_quarter,
            "_filter_anf_quarterly_segment_actual_rows": _filter_anf_quarterly_segment_actual_rows,
            "_parse_quarter_from_filename": _parse_quarter_from_filename,
            "_parse_quarter_from_follow_text": _parse_quarter_from_follow_text,
            "_path_belongs_to_ticker": _path_belongs_to_ticker,
            "_path_cache_key": _path_cache_key,
            "_read_cached_doc_raw": _read_cached_doc_raw,
            "_read_cached_doc_text": _read_cached_doc_text,
            "_infer_cached_doc_quarter": _infer_cached_doc_quarter,
            "_slide_text_paths": _slide_text_paths,
            "_resolve_col": _resolve_col,
            "_source_rank": _source_rank,
            "_text_fragment_penalty": _text_fragment_penalty,
            "_timed_writer_substage": _timed_writer_substage,
            "_parse_gpre_crush_margin_pair_local": _parse_gpre_crush_margin_pair_local,
            "_extract_45z_2026_target_candidates": _extract_45z_2026_target_candidates,
            "_extract_45z_monetization_target_display": _extract_45z_monetization_target_display,
            "_extract_money_targets_for_display": _extract_money_targets_for_display,
            "_anf_compact_driver_label": _anf_compact_driver_label,
            "_anf_visible_quarter_label": _anf_visible_quarter_label,
            "derivative_oci_bridge_df": derivative_oci_bridge_df_runtime,
        }

    def _get_operating_drivers_support() -> OperatingDriversSupport:
        nonlocal operating_drivers_support
        runtime = _operating_drivers_support_runtime()
        if operating_drivers_support is None:
            operating_drivers_support = OperatingDriversSupport(
                OperatingDriversSupportDeps(runtime=runtime)
            )
        else:
            operating_drivers_support.refresh_runtime(runtime)
        return operating_drivers_support

    def _sync_operating_drivers_support_cache_state() -> None:
        if operating_drivers_support is not None:
            operating_drivers_support.sync_runtime_cache_state()

    def _operating_driver_quarters() -> List[date]:
        return _get_operating_drivers_support().operating_driver_quarters()

    def _driver_source_display(source_type: Any, source_doc: Any = "") -> str:
        return _get_operating_drivers_support().driver_source_display(source_type, source_doc)

    def _driver_source_note(source_doc: Any, snippet: Any = "", extra: Any = "") -> str:
        return _get_operating_drivers_support().driver_source_note(source_doc, snippet, extra)

    cached_document_support: Optional[CachedDocumentSupport] = None

    def _cached_document_support_runtime() -> Dict[str, Any]:
        return {
            "document_cache": document_cache,
            "cache_roots": cache_roots,
            "cache_root": cache_root,
            "pdf_text_cache_root": pdf_text_cache_root,
            "rebuild_doc_text_cache": rebuild_doc_text_cache,
            "quiet_pdf_warnings": quiet_pdf_warnings,
            "_ticker_specific_submission_path": lambda path_in: _ticker_specific_submission_path(path_in),
            "source_path_cache_key": source_path_cache_key,
            "source_read_cached_doc_raw": source_read_cached_doc_raw,
            "source_read_cached_doc_text": source_read_cached_doc_text,
            "source_infer_cached_doc_quarter": source_infer_cached_doc_quarter,
            "source_sec_docs_for_accession": source_sec_docs_for_accession,
            "source_submission_cache_files": source_submission_cache_files,
            "source_submission_recent_rows": source_submission_recent_rows,
            "source_resolve_cached_doc_path": source_resolve_cached_doc_path,
            "_parse_quarter_from_filename": _parse_quarter_from_filename,
            "_parse_quarter_from_follow_text": _parse_quarter_from_follow_text,
            "parse_date": parse_date,
            "_is_quarter_end": _is_quarter_end,
            "_coerce_prev_quarter_end": _coerce_prev_quarter_end,
        }

    def _get_cached_document_support() -> CachedDocumentSupport:
        nonlocal cached_document_support
        if cached_document_support is None:
            cached_document_support = CachedDocumentSupport(
                CachedDocumentSupportDeps(runtime=_cached_document_support_runtime())
            )
        return cached_document_support

    def _path_cache_key(path_in: Path) -> str:
        return _get_cached_document_support().path_cache_key(path_in)

    def _read_cached_doc_raw(path_in: Path) -> str:
        return _get_cached_document_support().read_cached_doc_raw(path_in)

    def _read_cached_doc_text(path_in: Path, *, normalize: bool = False) -> str:
        return _get_cached_document_support().read_cached_doc_text(path_in, normalize=normalize)

    def _infer_cached_doc_quarter(
        path_in: Path,
        *,
        text: Any = None,
        latest_q_hint: Any = None,
        include_follow_text: bool = False,
    ) -> Optional[date]:
        return _get_cached_document_support().infer_cached_doc_quarter(
            path_in,
            text=text,
            latest_q_hint=latest_q_hint,
            include_follow_text=include_follow_text,
        )

    def _sec_docs_for_accession(accn_in: Any) -> List[Path]:
        return _get_cached_document_support().sec_docs_for_accession(accn_in)

    def _ticker_specific_submission_path(path_in: Path) -> bool:
        if not _path_belongs_to_ticker(path_in, ticker, ticker_roots):
            return False
        ticker_token = str(profile_ticker or ticker or "").strip().upper()
        if not ticker_token:
            return True
        try:
            path_parts = {str(part).strip().upper() for part in Path(path_in).resolve().parts}
        except Exception:
            path_parts = {str(part).strip().upper() for part in Path(path_in).parts}
        has_ticker_specific_root = False
        if cache_dir is not None:
            try:
                cache_parts = {str(part).strip().upper() for part in Path(cache_dir).expanduser().resolve().parts}
            except Exception:
                cache_parts = {str(part).strip().upper() for part in Path(cache_dir).parts}
            if ticker_token in cache_parts:
                has_ticker_specific_root = True
        for root_in in ticker_roots:
            try:
                root_parts = {str(part).strip().upper() for part in Path(root_in).resolve().parts}
            except Exception:
                root_parts = {str(part).strip().upper() for part in Path(root_in).parts}
            if ticker_token in root_parts:
                has_ticker_specific_root = True
                break
        if ticker_token in path_parts:
            return True
        # If no ticker-specific root is known, preserve the legacy broad-root
        # behavior; otherwise reject submissions from sibling ticker caches.
        return not has_ticker_specific_root

    def _submission_cache_files(*, max_files: Optional[int] = None) -> List[Path]:
        return _get_cached_document_support().submission_cache_files(max_files=max_files)

    def _submission_recent_row_quarter(row: Dict[str, Any]) -> Optional[date]:
        return _get_cached_document_support().submission_recent_row_quarter(row)

    def _submission_recent_rows(*, max_files: Optional[int] = None) -> List[Dict[str, Any]]:
        return _get_cached_document_support().submission_recent_rows(max_files=max_files)

    def _resolve_cached_doc_path(
        *,
        accn: Any = "",
        doc_name: Any = "",
        path_hint: Any = "",
    ) -> Optional[Path]:
        return _get_cached_document_support().resolve_cached_doc_path(
            accn=accn,
            doc_name=doc_name,
            path_hint=path_hint,
        )

    valuation_precompute_support = ValuationPrecomputeSupport(
        ValuationPrecomputeDeps(
            runtime={
                **globals(),
                **locals(),
            }
        )
    )
    _load_buyback_auth_source_bundle = valuation_precompute_support.load_buyback_auth_source_bundle
    _buyback_auth_docs_for_accession = valuation_precompute_support.buyback_auth_docs_for_accession
    _extract_latest_buyback_remaining_from_sec = valuation_precompute_support.extract_latest_buyback_remaining_from_sec
    _extract_valuation_filing_doc_text = valuation_precompute_support.extract_valuation_filing_doc_text
    _docs_for_valuation_accn = valuation_precompute_support.docs_for_valuation_accn
    _load_valuation_filing_docs_by_quarter = valuation_precompute_support.load_filing_docs_by_quarter
    _extract_cap_alloc_text_maps_by_quarter = valuation_precompute_support.extract_cap_alloc_text_maps_by_quarter
    _analyze_cap_alloc_doc = valuation_precompute_support.analyze_cap_alloc_doc
    _extract_buyback_dividend_from_doc_index = valuation_precompute_support.extract_buyback_dividend_from_doc_index
    _ensure_valuation_precompute_bundle = valuation_precompute_support.ensure_precompute_bundle

    def _slide_text_paths(
        *,
        kind: str = "all",
        quarter: Optional[date] = None,
    ) -> List[Path]:
        return source_slide_text_paths(
            material_roots=tuple(material_roots),
            document_cache=document_cache,
            parse_quarter_from_filename=_parse_quarter_from_filename,
            kind=kind,
            quarter=quarter,
        )

    def _read_operating_driver_text(path_in: Path) -> str:
        return _get_operating_drivers_support().read_operating_driver_text(path_in)

    def _operating_driver_follow_source_dirs() -> List[Tuple[str, Path]]:
        return _get_operating_drivers_support().operating_driver_follow_source_dirs()

    def _operating_driver_financial_statement_files() -> List[Path]:
        return _get_operating_drivers_support().operating_driver_financial_statement_files()

    def _load_operating_driver_source_records() -> List[Dict[str, Any]]:
        return _get_operating_drivers_support().load_source_records()

    def _load_operating_driver_source_records_by_quarter() -> Dict[date, List[Dict[str, Any]]]:
        return _get_operating_drivers_support().load_source_records_by_quarter()

    def _load_operating_driver_line_index_by_quarter() -> Dict[date, List[Dict[str, Any]]]:
        return _get_operating_drivers_support().load_line_index_by_quarter()

    def _load_operating_driver_flat_line_index() -> List[Dict[str, Any]]:
        return _get_operating_drivers_support().load_flat_line_index()

    def _is_crush_margin_bridge_candidate(text_in: Any) -> bool:
        return _get_operating_drivers_support().is_crush_margin_bridge_candidate(text_in)

    def _parse_driver_number(token: Any) -> Optional[float]:
        return _get_operating_drivers_support().parse_driver_number(token)

    def _extract_driver_numeric_values(text_in: Any) -> List[float]:
        return _get_operating_drivers_support().extract_driver_numeric_values(text_in)

    def _get_crush_margin_bridge_details(text_in: Any) -> Dict[str, Any]:
        return _get_operating_drivers_support().get_crush_margin_bridge_details(text_in)

    def _prime_operating_driver_crush_detail_cache(records: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Dict[str, Any]]:
        return _get_operating_drivers_support().prime_operating_driver_crush_detail_cache(records)

    def _load_operating_driver_template_index() -> Dict[str, Any]:
        return _get_operating_drivers_support().load_template_index()

    def _load_operating_driver_45z_guidance_docs_by_quarter() -> Dict[date, List[Dict[str, Any]]]:
        return _get_operating_drivers_support().load_45z_guidance_docs_by_quarter()

    def _operating_driver_template_spec(tpl: Any) -> Dict[str, Any]:
        return _get_operating_drivers_support().operating_driver_template_spec(tpl)

    def _candidate_records_for_template(
        qd: date,
        template_spec: Dict[str, Any],
        quarter_records: Optional[List[Dict[str, Any]]] = None,
    ) -> List[Dict[str, Any]]:
        return _get_operating_drivers_support().candidate_records_for_template(
            qd,
            template_spec,
            quarter_records=quarter_records,
        )

    def _load_operating_driver_bridge_bundle_map(quarter_set: List[date]) -> Dict[date, Dict[str, Any]]:
        return _get_operating_drivers_support().load_bridge_bundle_map(quarter_set)

    def _get_analysis_sheet_style_bundle() -> Dict[str, Any]:
        return _style_get_analysis_sheet_style_bundle(header_size=header_size, font_size=font_size)

    def _get_valuation_style_bundle() -> Dict[str, Any]:
        nonlocal valuation_style_bundle_cache
        runtime: Dict[str, Any] = {
            "wb": wb,
            "ctx_ref": ctx_ref,
            "valuation_style_bundle_cache": valuation_style_bundle_cache,
            "copy": copy,
            "PatternFill": PatternFill,
            "get_column_letter": get_column_letter,
            "font_size": font_size,
            "_get_analysis_sheet_style_bundle": _get_analysis_sheet_style_bundle,
            "_timed_writer_substage": _timed_writer_substage,
        }
        bundle = get_valuation_style_bundle(ValuationStyleBundleDeps(runtime=runtime))
        valuation_style_bundle_cache = runtime.get("valuation_style_bundle_cache")
        return bundle

    def _ensure_valuation_render_bundle(qs_local: Tuple[pd.Timestamp, ...], leverage_df_local: Optional[pd.DataFrame]) -> Dict[str, Any]:
        nonlocal valuation_render_bundle_cache
        runtime: Dict[str, Any] = {
            "pd": pd,
            "hist": hist,
            "ctx_ref": ctx_ref,
            "valuation_render_bundle_cache": valuation_render_bundle_cache,
            "_hist_view": _hist_view,
            "_timed_writer_substage": _timed_writer_substage,
            "_shared_load_local_balance_sheet_detail_payloads": _shared_load_local_balance_sheet_detail_payloads,
            "_carry_forward_low_change_series": _carry_forward_low_change_series,
        }
        bundle = ensure_valuation_render_bundle(
            ValuationRenderBundleDeps(runtime=runtime),
            qs_local,
            leverage_df_local,
        )
        valuation_render_bundle_cache = runtime.get("valuation_render_bundle_cache")
        return bundle

    def _parse_threshold_amount_m(text_in: Any) -> Optional[float]:
        return _get_operating_drivers_support().parse_threshold_amount_m(text_in)

    def _parse_45z_realized_value_m(text_in: Any) -> Optional[float]:
        return _get_operating_drivers_support().parse_45z_realized_value_m(text_in)

    def _merge_driver_rows(existing: Dict[str, Any], candidate: Dict[str, Any]) -> Dict[str, Any]:
        return _get_operating_drivers_support().merge_driver_rows(existing, candidate)

    def _make_driver_row(*args: Any, **kwargs: Any) -> Dict[str, Any]:
        return _get_operating_drivers_support().make_driver_row(*args, **kwargs)

    def _build_anf_operating_driver_rows() -> List[Dict[str, Any]]:
        return _get_operating_drivers_support().build_anf_operating_driver_rows()

    def _gpre_canonical_crush_series_for_drivers_local() -> Dict[date, Dict[str, Any]]:
        return _get_operating_drivers_support().gpre_canonical_crush_series_for_drivers_local()

    def _extract_operating_driver_rows_for_template(
        qd: date,
        tpl: Any,
        *,
        quarter_records: Optional[List[Dict[str, Any]]] = None,
    ) -> List[Dict[str, Any]]:
        return _get_operating_drivers_support().extract_rows_for_template(
            qd,
            tpl,
            quarter_records=quarter_records,
        )

    def _format_operating_driver_delta(current_val: Any, prior_val: Any, unit: str) -> str:
        return _get_operating_drivers_support().format_operating_driver_delta(current_val, prior_val, unit)

    def _build_operating_drivers_history_rows() -> List[Dict[str, Any]]:
        return _get_operating_drivers_support().build_operating_drivers_history_rows()

    def _driver_unit_label(unit_txt: Any) -> str:
        return _get_operating_drivers_support().driver_unit_label(unit_txt)

    def _driver_row_label(driver_label: Any, unit_txt: Any) -> str:
        return _get_operating_drivers_support().driver_row_label(driver_label, unit_txt)

    def _truncate_driver_text(txt: Any, max_chars: int = 96) -> str:
        return _get_operating_drivers_support().truncate_driver_text(txt, max_chars)

    def _operating_driver_order_map(templates_in: List[Any]) -> Dict[str, int]:
        return _get_operating_drivers_support().operating_driver_order_map(templates_in)

    def _quarter_label_short(qd: Optional[date]) -> str:
        if not isinstance(qd, date):
            return ""
        if is_anf_profile:
            label = _anf_visible_quarter_label(qd)
            if label:
                return label
        return f"{qd.year}-Q{((qd.month - 1) // 3) + 1}"

    def _build_economics_market_rows() -> List[Dict[str, Any]]:
        deps = EconomicsMarketRowsDeps(
            cache_dir=cache_dir,
            ticker=ticker,
            company_profile=company_profile,
            first_existing_material_dir=_first_existing_material_dir,
            load_market_export_rows=load_market_export_rows,
        )
        return build_economics_market_rows(deps)

    def _write_operating_drivers_raw_sheet(rows: List[Dict[str, Any]]) -> None:
        deps = OperatingDriversRawSheetDeps(
            runtime={
                "wb": wb,
                "pd": pd,
                "ILLEGAL_CHARACTERS_RE": ILLEGAL_CHARACTERS_RE,
                "PatternFill": PatternFill,
                "Border": Border,
                "Side": Side,
                "Font": Font,
                "Alignment": Alignment,
                "header_size": header_size,
                "_safe_cell": _safe_cell,
                "_set_cell_comment_local": _set_cell_comment_local,
                "_estimate_wrapped_row_height": _estimate_wrapped_row_height,
            }
        )
        OperatingDriversRawSheetWriter(deps).write_operating_drivers_raw_sheet(rows)

    def _write_economics_market_raw_sheet(rows: List[Dict[str, Any]]) -> None:
        deps = EconomicsMarketRawWriterDeps(
            wb=wb,
            header_size=header_size,
            safe_cell=_safe_cell,
            estimate_wrapped_row_height=_estimate_wrapped_row_height,
        )
        write_economics_market_raw_sheet(deps, rows)

    def _write_operating_drivers_sheet(rows: List[Dict[str, Any]]) -> None:
        deps = OperatingDriversSheetAdapterDeps(
            runtime={
                "OperatingDriversWriterDeps": OperatingDriversWriterDeps,
                "write_operating_drivers_sheet": write_operating_drivers_sheet,
                "wb": wb,
                "hist": hist,
                "ticker": ticker,
                "company_profile": company_profile,
                "slides_segments": slides_segments,
                "slides_guidance": slides_guidance,
                "quarter_notes": quarter_notes,
                "derivative_oci_bridge_df": derivative_oci_bridge_df,
                "material_roots": material_roots,
                "font_size": font_size,
                "header_size": header_size,
                "is_pbi_profile": is_pbi_profile,
                "is_gpre_profile": is_gpre_profile,
                "is_anf_profile": is_anf_profile,
                "enable_quarterly_segment_block": enable_quarterly_segment_block,
                "annual_segment_alias_patterns": annual_segment_alias_patterns,
                "ANF_SEGMENT_BRAND_EXPLANATION": ANF_SEGMENT_BRAND_EXPLANATION,
                "_get_valuation_style_bundle": _get_valuation_style_bundle,
                "_get_analysis_sheet_style_bundle": _get_analysis_sheet_style_bundle,
                "_operating_driver_quarters": _operating_driver_quarters,
                "_load_operating_driver_template_index": _load_operating_driver_template_index,
                "_load_operating_driver_source_records_by_quarter": _load_operating_driver_source_records_by_quarter,
                "_load_operating_driver_flat_line_index": _load_operating_driver_flat_line_index,
                "_first_existing_material_dir": _first_existing_material_dir,
                "_parse_quarter_from_filename": _parse_quarter_from_filename,
                "_parse_quarter_from_follow_text": _parse_quarter_from_follow_text,
                "_read_operating_driver_text": _read_operating_driver_text,
                "_set_cell_comment_local": _set_cell_comment_local,
                "_driver_source_note": _driver_source_note,
                "_driver_source_display": _driver_source_display,
                "_driver_row_label": _driver_row_label,
                "_truncate_driver_text": _truncate_driver_text,
                "_quarter_label_short": _quarter_label_short,
                "_source_rank": _source_rank,
                "_text_fragment_penalty": _text_fragment_penalty,
                "_ensure_terminal_period": _ensure_terminal_period,
                "_gpre_commercial_setup_records_shared": _gpre_commercial_setup_records_shared,
                "_anf_clean_visible_operating_driver_records": _anf_clean_visible_operating_driver_records,
                "_anf_clean_visible_ui_text": _anf_clean_visible_ui_text,
                "_anf_compact_driver_group": _anf_compact_driver_group,
                "_anf_compact_driver_label": _anf_compact_driver_label,
                "_anf_recent_operating_commentary_rows": _anf_recent_operating_commentary_rows,
                "_anf_round_visible_driver_value": _anf_round_visible_driver_value,
                "_anf_visible_quarter_label": _anf_visible_quarter_label,
                "_sector_operating_driver_intro_tables": _sector_operating_driver_intro_tables,
            }
        )
        OperatingDriversSheetAdapter(deps).write_operating_drivers_sheet(rows)



    def _economics_overlay_sheet_runtime() -> Dict[str, Any]:
        return {
            "BasisProxySandboxWriterDeps": BasisProxySandboxWriterDeps,
            "EconomicsOverlayChartWriterDeps": EconomicsOverlayChartWriterDeps,
            "EconomicsOverlayMarketStateDeps": EconomicsOverlayMarketStateDeps,
            "EconomicsOverlaySourceSupport": EconomicsOverlaySourceSupport,
            "EconomicsOverlaySourceSupportDeps": EconomicsOverlaySourceSupportDeps,
            "GpreEconomicsOverlayBridgeDeps": GpreEconomicsOverlayBridgeDeps,
            "GpreEconomicsOverlayCommercialDeps": GpreEconomicsOverlayCommercialDeps,
            "GpreEconomicsOverlayCoproductDeps": GpreEconomicsOverlayCoproductDeps,
            "GpreEconomicsOverlayCurrentQtdDeps": GpreEconomicsOverlayCurrentQtdDeps,
            "GpreEconomicsOverlayDerivativeSideEffectDeps": GpreEconomicsOverlayDerivativeSideEffectDeps,
            "GpreEconomicsOverlayInputRowsDeps": GpreEconomicsOverlayInputRowsDeps,
            "GpreOverlayQuarterComparisonDeps": GpreOverlayQuarterComparisonDeps,
            "GpreOverlaySupportInputs": GpreOverlaySupportInputs,
            "_apply_chart_text_categories": _apply_chart_text_categories,
            "_convert_market_price_value": _convert_market_price_value,
            "_driver_source_display": _driver_source_display,
            "_driver_source_note": _driver_source_note,
            "_economics_market_region_tags": _economics_market_region_tags,
            "_ensure_terminal_period": _ensure_terminal_period,
            "_estimate_wrapped_row_height": _estimate_wrapped_row_height,
            "_extract_operating_driver_rows_for_template": _extract_operating_driver_rows_for_template,
            "_get_analysis_sheet_style_bundle": _get_analysis_sheet_style_bundle,
            "_gpre_commercial_setup_records_shared": _gpre_commercial_setup_records_shared,
            "_gpre_parse_snapshot_date_like": _gpre_parse_snapshot_date_like,
            "_load_operating_driver_bridge_bundle_map": _load_operating_driver_bridge_bundle_map,
            "_load_operating_driver_flat_line_index": _load_operating_driver_flat_line_index,
            "_load_operating_driver_source_records_by_quarter": _load_operating_driver_source_records_by_quarter,
            "_load_operating_driver_template_index": _load_operating_driver_template_index,
            "_operating_driver_quarters": _operating_driver_quarters,
            "_overlay_model_label": _overlay_model_label,
            "_parse_driver_number": _parse_driver_number,
            "_quarter_label_short": _quarter_label_short,
            "_record_writer_substage": _record_writer_substage,
            "_set_cell_comment_local": _set_cell_comment_local,
            "_text_fragment_penalty": _text_fragment_penalty,
            "_truncate_driver_text": _truncate_driver_text,
            "_write_derivative_crush_tests_sheet": _write_derivative_crush_tests_sheet,
            "build_current_qtd_simple_crush_snapshot": build_current_qtd_simple_crush_snapshot,
            "build_derivative_crush_tests": build_derivative_crush_tests,
            "build_economics_overlay_market_state": build_economics_overlay_market_state,
            "build_gpre_basis_proxy_model": build_gpre_basis_proxy_model,
            "build_gpre_official_proxy_history_series": build_gpre_official_proxy_history_series,
            "build_gpre_official_proxy_snapshot": build_gpre_official_proxy_snapshot,
            "build_gpre_overlay_proxy_preview_bundle": build_gpre_overlay_proxy_preview_bundle,
            "build_gpre_plant_capacity_history": build_gpre_plant_capacity_history,
            "build_next_quarter_thesis_snapshot": build_next_quarter_thesis_snapshot,
            "build_prior_quarter_simple_crush_snapshot": build_prior_quarter_simple_crush_snapshot,
            "build_simple_crush_history_series": build_simple_crush_history_series,
            "cache_dir": cache_dir,
            "company_profile": company_profile,
            "data_root_from_sec_cache_path": data_root_from_sec_cache_path,
            "derivative_oci_bridge_df": derivative_oci_bridge_df,
            "derivative_oci_exposure_df": derivative_oci_exposure_df,
            "economics_market_rows": economics_market_rows,
            "fetch_gpre_corn_bids_snapshot": fetch_gpre_corn_bids_snapshot,
            "font_size": font_size,
            "glx_normalize_text": glx_normalize_text,
            "header_size": header_size,
            "info_log": info_log,
            "is_gpre_profile": is_gpre_profile,
            "is_pbi_profile": is_pbi_profile,
            "load_or_download_gpre_corn_bids_snapshot": load_or_download_gpre_corn_bids_snapshot,
            "market_build_gpre_proxy_implied_results_bundle": market_build_gpre_proxy_implied_results_bundle,
            "market_gpre_phase_preview_story": market_gpre_phase_preview_story,
            "market_input_fingerprint": market_input_fingerprint,
            "operating_driver_history_rows": operating_driver_history_rows,
            "persist_gpre_frozen_thesis_snapshot": persist_gpre_frozen_thesis_snapshot,
            "qn_is_complete_signal_text": qn_is_complete_signal_text,
            "resolve_gpre_quarter_open_snapshot": resolve_gpre_quarter_open_snapshot,
            "state": state,
            "ticker": ticker,
            "ticker_roots": ticker_roots,
            "wb": wb,
            "write_basis_proxy_sandbox_sheet": write_basis_proxy_sandbox_sheet,
            "write_economics_overlay_charts": write_economics_overlay_charts,
            "write_gpre_basis_proxy_overlay_support": write_gpre_basis_proxy_overlay_support,
            "write_gpre_derivative_crush_tests_side_effect": write_gpre_derivative_crush_tests_side_effect,
            "write_gpre_economics_overlay_bridge_to_reported_section": write_gpre_economics_overlay_bridge_to_reported_section,
            "write_gpre_economics_overlay_commercial_sections": write_gpre_economics_overlay_commercial_sections,
            "write_gpre_economics_overlay_coproduct_section": write_gpre_economics_overlay_coproduct_section,
            "write_gpre_economics_overlay_current_qtd_section": write_gpre_economics_overlay_current_qtd_section,
            "write_gpre_economics_overlay_input_rows": write_gpre_economics_overlay_input_rows,
            "write_gpre_overlay_quarter_comparisons": write_gpre_overlay_quarter_comparisons,
        }

    def _write_economics_overlay_sheet(rows: List[Dict[str, Any]]) -> None:
        deps = EconomicsOverlaySheetDeps(runtime=_economics_overlay_sheet_runtime())
        return EconomicsOverlaySheetWriter(deps).write_economics_overlay_sheet(rows)

    def _write_anf_investment_case_surfaces() -> pd.DataFrame:
        if not is_anf_profile:
            return pd.DataFrame()
        guidance_for_case = slides_guidance
        try:
            guidance_for_case = _anf_visible_guidance_normalized_frame(guidance_for_case)
        except Exception:
            guidance_for_case = slides_guidance
        case_data = _anf_build_investment_case_data(
            hist=hist,
            operating_driver_rows=operating_driver_history_rows,
            guidance_normalized=guidance_for_case,
            slides_segments=slides_segments,
            valuation_summary=valuation_summary_df,
            adjusted_metrics=adj_metrics,
        )
        _write_anf_investment_case_sheet(wb, case_data)
        _write_anf_investment_case_data_sheet(wb, case_data)
        return case_data

    def _write_investment_case_surfaces() -> pd.DataFrame:
        ticker_txt = str(ticker or "").strip().upper()
        if is_anf_profile:
            return _write_anf_investment_case_surfaces()
        if ticker_txt not in {"PBI", "GPRE"}:
            return pd.DataFrame()
        guidance_for_case = slides_guidance
        try:
            from .excel_writer_ui import _shared_guidance_normalized_frame

            guidance_for_case = _shared_guidance_normalized_frame(guidance_for_case)
        except Exception:
            guidance_for_case = slides_guidance
        case_data = _sector_build_investment_case_data(
            ticker=ticker_txt,
            hist=hist,
            operating_driver_rows=operating_driver_history_rows,
            guidance_normalized=guidance_for_case,
            valuation_summary=valuation_summary_df,
            economics_market_rows=economics_market_rows,
            slides_segments=slides_segments,
        )
        _write_sector_investment_case_sheet(wb, ticker_txt, case_data)
        _write_sector_investment_case_data_sheet(wb, ticker_txt, case_data)
        return case_data

    def _write_quarter_narrative_data_surface() -> None:
        return _write_quarter_narrative_data_surface_impl(_quarter_notes_context_adapter_deps())

    def _write_quarter_notes_narrative_ui_surface() -> None:
        return _write_quarter_notes_narrative_ui_surface_impl(_quarter_notes_context_adapter_deps())

    enable_derivative_oci_bridge_sheet = bool(
        (is_gpre_profile or bool(getattr(company_profile, "enable_derivative_oci_bridge", False)))
        and not is_pbi_profile
    )
    # The accounting bridge belongs next to Promise_Progress_UI, while the
    # crush-testing surface sits after Basis_Proxy_Sandbox because it consumes
    # that sheet's market/proxy quarterly frame. PBI stays excluded until it has
    # an explicit derivative/OCI bridge use case.
    derivative_sheet_order_slot = ("Derivative_OCI_Bridge",) if enable_derivative_oci_bridge_sheet else tuple()
    derivative_crush_tests_order_slot = ("Derivative_Crush_Tests",) if (enable_derivative_oci_bridge_sheet and is_gpre_profile) else tuple()
    desired_sheet_order = (
        "SUMMARY",
        "Valuation",
        "BS_Segments",
        "Operating_Drivers",
        "Economics_Overlay",
        "Quarter_Notes_UI",
        "Promise_Progress_UI",
        *derivative_sheet_order_slot,
        "Basis_Proxy_Sandbox",
        *derivative_crush_tests_order_slot,
        "Hidden_Value_Flags",
        "Revolver_History",
        "Debt_Tranches_Latest",
        "Debt_Profile",
        "Debt_Maturity_Ladder",
        "Debt_Buckets",
        "Debt_Recon",
        "Debt_Tranches_Q",
        "Debt_Credit_Notes",
        "Leverage_Liquidity",
        "REPORT_IS_Q",
        "REPORT_BS_Q",
        "REPORT_CF_Q",
        "Quarter_Notes",
        "Quarter_Notes_Evidence",
        "Quarter_Narrative_Data",
        "Quarter_Notes_Audit",
    )
    raw_sheet_cluster = ("History_Q", "operating_drivers_raw", "economics_market_raw")
    desired_sheet_order, raw_sheet_cluster = _investment_case_sheet_order(
        desired_sheet_order,
        raw_sheet_cluster,
        ticker=ticker,
    )
    runtime_cache.valuation_style_bundle_cache = valuation_style_bundle_cache
    runtime_cache.valuation_render_bundle_cache = valuation_render_bundle_cache
    runtime_cache.valuation_precompute_bundle_cache = valuation_precompute_support.valuation_precompute_bundle_cache
    runtime_cache.valuation_filing_docs_by_quarter_cache = valuation_precompute_support.valuation_filing_docs_by_quarter_cache
    _sync_operating_drivers_support_cache_state()
    _refresh_profile_signal_runtime({**globals(), **locals()})
    operating_drivers_runtime.profile_slide_signals_cache = profile_slide_signals_cache
    operating_drivers_runtime.profile_slide_signals_by_quarter_cache = profile_slide_signals_by_quarter_cache
    runtime_cache.adj_net_leverage_text_map_cache = adj_net_leverage_text_map_cache
    runtime_cache.leverage_local_material_index_cache = leverage_local_material_index_cache
    runtime_cache.leverage_audit_doc_index_cache = leverage_audit_doc_index_cache
    runtime_cache.promise_progress_ui_bundle_cache = promise_progress_ui_bundle_cache
    runtime_cache.valuation_buyback_auth_source_bundle_cache = valuation_precompute_support.valuation_buyback_auth_source_bundle_cache
    derivative_oci_bridge_df = pd.DataFrame()
    derivative_oci_qa_df = pd.DataFrame()
    derivative_oci_exposure_df = pd.DataFrame()
    if enable_derivative_oci_bridge_sheet:
        try:
            derivative_oci_result = build_derivative_oci_bridge_from_sources(
                str(ticker or ""),
                hist=hist if isinstance(hist, pd.DataFrame) else None,
                adj_metrics=adj_metrics if isinstance(adj_metrics, pd.DataFrame) else None,
            )
            derivative_oci_bridge_df = derivative_oci_result.rows
            derivative_oci_qa_df = derivative_oci_result.qa_rows
            derivative_oci_exposure_df = derivative_oci_result.exposure_rows
        except Exception as exc:
            derivative_oci_qa_df = pd.DataFrame(
                [
                    {
                        "quarter": pd.NaT,
                        "metric": "Derivative & OCI Bridge",
                        "severity": "warn",
                        "message": f"Derivative/OCI bridge extraction failed; workbook actuals were left unchanged. Error: {exc}",
                        "source": "derivative_oci_bridge",
                        "issue_family": "derivative_disclosure_method",
                        "recommended_action": "review derivative/OCI source parsing",
                        "raw_metric": "derivative_notes",
                    }
                ]
            )
    runtime_data = WriterRuntimeData(
        out_path=out_path,
        ticker=ticker,
        excel_mode=excel_mode,
        profile_timings=profile_timings,
        quarter_notes_audit=quarter_notes_audit,
        enable_operating_drivers_sheet=enable_operating_drivers_sheet,
        enable_economics_overlay_sheet=enable_economics_overlay_sheet,
        enable_economics_market_raw_sheet=enable_economics_market_raw_sheet,
        operating_driver_history_rows=operating_driver_history_rows,
        economics_market_rows=economics_market_rows,
        qa_checks=qa_checks if isinstance(qa_checks, pd.DataFrame) else pd.DataFrame(),
        info_log=info_log if isinstance(info_log, pd.DataFrame) else pd.DataFrame(),
        data_is_rules_df=data_is_rules_df if isinstance(data_is_rules_df, pd.DataFrame) else pd.DataFrame(),
        doc_cache=document_cache,
        frame_view_cache=frame_view_cache,
        runtime_cache=runtime_cache,
        extra_values={
            "leverage_df": leverage_df,
            "valuation_summary_df": valuation_summary_df,
            "valuation_grid_df": valuation_grid_df,
            "summary_df": summary_df,
            "signals_base_df": signals_base_df,
            "flags_df": flags_df,
            "flags_audit_df": flags_audit_df,
            "flags_recompute_df": flags_recompute_df,
            "ng_bridge": ng_bridge,
            "ng_bridge_relaxed": ng_bridge_relaxed,
            "report_is": pd.DataFrame(),
            "report_bs": pd.DataFrame(),
            "report_cf": pd.DataFrame(),
            "facts_long": pd.DataFrame(),
            "lineitem_map": pd.DataFrame(),
            "period_index": pd.DataFrame(),
            "quarter_notes_evidence_df": pd.DataFrame(),
            "promise_evidence_df": pd.DataFrame(),
            "derivative_oci_bridge_df": derivative_oci_bridge_df,
            "derivative_oci_qa_df": derivative_oci_qa_df,
            "derivative_oci_exposure_df": derivative_oci_exposure_df,
            "company_profile": company_profile,
            "font_size": font_size,
            "header_size": header_size,
        },
    )
    callbacks = WriterCallbacks(
        write_sheet=_write_sheet,
        write_flags_sheet=_write_flags_sheet,
        write_report_sheet=_write_report_sheet,
        write_summary_sheet=_write_summary_sheet,
        write_valuation_sheet=_write_valuation_sheet,
        write_bs_segments_sheet=_write_bs_segments_sheet,
        write_quarter_notes_ui_v2=_write_quarter_notes_ui_v2,
        write_promise_tracker_ui_v2=_write_promise_tracker_ui_v2,
        write_promise_progress_ui_v2=_write_promise_progress_ui_v2,
        write_operating_drivers_sheet=_write_operating_drivers_sheet,
        write_economics_overlay_sheet=_write_economics_overlay_sheet,
        write_operating_drivers_raw_sheet=_write_operating_drivers_raw_sheet,
        write_economics_market_raw_sheet=_write_economics_market_raw_sheet,
        build_report=_build_report,
        build_summary=_build_summary,
        build_facts_long=_build_facts_long,
        build_lineitem_map=_build_lineitem_map,
        build_period_index=_build_period_index,
        build_ng_bridge=_build_ng_bridge,
        build_qn_evidence_src=_build_qn_evidence_src,
        build_promise_evidence_src=_build_promise_evidence_src,
        extract_adj_net_leverage_text_map=_extract_adj_net_leverage_text_map,
        build_hidden_value_flags_fallback=_build_hidden_value_flags_fallback,
        load_operating_driver_source_records=_load_operating_driver_source_records,
        load_operating_driver_source_records_by_quarter=_load_operating_driver_source_records_by_quarter,
        prime_operating_driver_crush_detail_cache=_prime_operating_driver_crush_detail_cache,
        build_operating_drivers_history_rows=_build_operating_drivers_history_rows,
        build_economics_market_rows=_build_economics_market_rows,
        run_latest_quarter_qa=_run_latest_quarter_qa,
        extra_callbacks={
            "_ui_state": ui_state,
            "_slide_text_paths": _slide_text_paths,
            "_pbi_slide_pages_for_qd": _pbi_slide_pages_for_qd,
            "_latest_quarter_qa_source_bundle": _latest_quarter_qa_source_bundle,
            "_latest_quarter_sec_text_corpus": _latest_quarter_sec_text_corpus,
            "_local_slide_driver_fallback": _local_slide_driver_fallback,
            "_load_profile_slide_signals": _load_profile_slide_signals,
            "_quarter_notes_view": _quarter_notes_view,
            "_promises_view": _promises_view,
            "_audit_view": _audit_view,
            "_hist_view": _hist_view,
            "_adj_metrics_view": _adj_metrics_view,
            "_load_operating_driver_line_index_by_quarter": _load_operating_driver_line_index_by_quarter,
            "_load_operating_driver_flat_line_index": _load_operating_driver_flat_line_index,
            "_load_profile_slide_signals_by_quarter": _load_profile_slide_signals_by_quarter,
            "_load_operating_driver_45z_guidance_docs_by_quarter": _load_operating_driver_45z_guidance_docs_by_quarter,
            "_ensure_valuation_render_bundle": _ensure_valuation_render_bundle,
            "_ensure_valuation_precompute_bundle": _ensure_valuation_precompute_bundle,
            "_write_derivative_oci_bridge_sheet": _write_derivative_oci_bridge_sheet,
            "_write_investment_case_surfaces": _write_investment_case_surfaces,
            "_write_anf_investment_case_surfaces": _write_anf_investment_case_surfaces,
            "_write_quarter_notes_narrative_ui_sheet": _write_quarter_notes_narrative_ui_surface,
            "_write_quarter_narrative_data_sheet": _write_quarter_narrative_data_surface,
            "_apply_shared_ui_conventions": lambda: _apply_shared_ui_conventions_to_workbook(wb, ticker),
            "_final_promise_progress_cleanup": lambda: _final_repair_promise_progress_ui(wb, ticker),
        },
    )
    _quarter_notes_ui_selection_outer_scope.update(locals())
    ctx = WriterContext(
        inputs=inputs,
        wb=wb,
        font_size=font_size,
        header_size=header_size,
        company_profile=company_profile,
        data=runtime_data,
        callbacks=callbacks,
        writer_timings={},
        ui_info_rows=ui_info_rows,
        desired_sheet_order=desired_sheet_order,
        raw_sheet_cluster=raw_sheet_cluster,
    )
    ctx_ref = ctx
    valuation_precompute_support.runtime["ctx_ref"] = ctx_ref
    state = _build_compat_state(ctx)
    ctx.state = state
    return ctx
