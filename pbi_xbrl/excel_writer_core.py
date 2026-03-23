from __future__ import annotations

import json
import re
import time
from bisect import bisect_right
from contextlib import contextmanager
from typing import Any, Dict, Iterator, List, Optional, Tuple

import pandas as pd
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from .excel_writer_context import WriterContext
from .signals import build_hidden_value_outputs, build_signals_base
from .valuation import valuation_engine, valuation_to_frames


@contextmanager
def timed_writer_stage(
    writer_timings: Dict[str, float],
    name: str,
    *,
    enabled: bool,
) -> Iterator[None]:
    start = time.perf_counter()
    try:
        yield
    finally:
        dt_s = time.perf_counter() - start
        writer_timings[name] = writer_timings.get(name, 0.0) + dt_s
        if enabled:
            print(f"[timing] {name}={dt_s:.2f}s", flush=True)


def _sync_frame_placeholder(holder: Any, value: Optional[pd.DataFrame]) -> pd.DataFrame:
    replacement = value.copy() if isinstance(value, pd.DataFrame) else pd.DataFrame()
    if isinstance(holder, pd.DataFrame) and hasattr(holder, "_update_inplace"):
        holder._update_inplace(replacement)
        return holder
    return replacement


def ensure_driver_inputs(ctx: WriterContext) -> None:
    data = ctx.data
    callbacks = ctx.callbacks
    state = ctx.state
    if data.driver_inputs_ready:
        return
    with timed_writer_stage(
        ctx.writer_timings,
        "write_excel.derive.driver_inputs",
        enabled=bool(ctx.inputs.profile_timings),
    ):
        if data.enable_operating_drivers_sheet:
            with timed_writer_stage(
                ctx.writer_timings,
                "write_excel.derive.driver_inputs.source_records",
                enabled=bool(ctx.inputs.profile_timings),
            ):
                callbacks.load_operating_driver_source_records()
                callbacks.load_operating_driver_source_records_by_quarter()
            line_index_loader = callbacks.extra_callbacks.get("_load_operating_driver_line_index_by_quarter")
            if callable(line_index_loader):
                with timed_writer_stage(
                    ctx.writer_timings,
                    "write_excel.derive.driver_inputs.line_index",
                    enabled=bool(ctx.inputs.profile_timings),
                ):
                    line_index_loader()
            with timed_writer_stage(
                ctx.writer_timings,
                "write_excel.derive.driver_inputs.crush_bridge_cache",
                enabled=bool(ctx.inputs.profile_timings),
            ):
                callbacks.prime_operating_driver_crush_detail_cache()
            with timed_writer_stage(
                ctx.writer_timings,
                "write_excel.derive.driver_inputs.operating_history",
                enabled=bool(ctx.inputs.profile_timings),
            ):
                operating_driver_history_rows = callbacks.build_operating_drivers_history_rows()
        else:
            operating_driver_history_rows = []
        data.operating_driver_history_rows = list(operating_driver_history_rows)
        state["operating_driver_history_rows"] = data.operating_driver_history_rows

        if data.enable_economics_overlay_sheet or data.enable_economics_market_raw_sheet:
            with timed_writer_stage(
                ctx.writer_timings,
                "write_excel.derive.driver_inputs.economics_market",
                enabled=bool(ctx.inputs.profile_timings),
            ):
                economics_market_rows = callbacks.build_economics_market_rows()
        else:
            economics_market_rows = []
        shared_market_rows = data.economics_market_rows
        if isinstance(shared_market_rows, list):
            shared_market_rows[:] = list(economics_market_rows)
            data.economics_market_rows = shared_market_rows
        else:
            data.economics_market_rows = list(economics_market_rows)
        state["economics_market_rows"] = data.economics_market_rows
        data.driver_inputs_ready = True
        state["_driver_inputs_ready"] = True


def prepare_writer_inputs(ctx: WriterContext) -> None:
    ctx.data.driver_inputs_ready = False
    ctx.state["_driver_inputs_ready"] = False


def ensure_report_inputs(ctx: WriterContext) -> None:
    callbacks = ctx.callbacks
    state = ctx.state
    if ctx.derived.report_is is not None:
        return
    with timed_writer_stage(
        ctx.writer_timings,
        "write_excel.derive.report_inputs",
        enabled=bool(ctx.inputs.profile_timings),
    ):
        report_is = _sync_frame_placeholder(state.get("report_is"), callbacks.build_report("IS", scale=1e6))
        report_bs = _sync_frame_placeholder(state.get("report_bs"), callbacks.build_report("BS", scale=1e6))
        report_cf = _sync_frame_placeholder(state.get("report_cf"), callbacks.build_report("CF", scale=1e6))
        state["report_is"] = report_is
        state["report_bs"] = report_bs
        state["report_cf"] = report_cf
        ctx.derived.report_is = report_is
        ctx.derived.report_bs = report_bs
        ctx.derived.report_cf = report_cf


def _ensure_valuation_source_views(ctx: WriterContext) -> None:
    derived = ctx.derived
    if derived.valuation_hist_view is not None:
        return

    hist = ctx.inputs.hist
    revolver_history = ctx.inputs.revolver_history
    revolver_df = ctx.inputs.revolver_df
    adj_metrics = ctx.inputs.adj_metrics
    with timed_writer_stage(
        ctx.writer_timings,
        "write_excel.derive.valuation_inputs.normalize_sources",
        enabled=bool(ctx.inputs.profile_timings),
    ):
        hist_view = pd.DataFrame()
        hist_indexed = pd.DataFrame()
        latest_context: Dict[str, Any] = {}
        last4_context: Dict[str, Any] = {}
        core_maps: Dict[str, Any] = {}
        if hist is not None and not hist.empty and "quarter" in hist.columns:
            hist_view = hist.copy()
            hist_view["quarter"] = pd.to_datetime(hist_view["quarter"], errors="coerce")
            hist_view = hist_view[hist_view["quarter"].notna()].sort_values("quarter").reset_index(drop=True)
            if not hist_view.empty:
                hist_indexed = hist_view.drop_duplicates(subset=["quarter"], keep="last").set_index("quarter")

                quarter_list = [pd.Timestamp(q) for q in hist_view["quarter"].tolist()]
                numeric_cols: Dict[str, pd.Series] = {}
                for col in (
                    "cash",
                    "debt_core",
                    "ebitda",
                    "interest_paid",
                    "interest_expense_net",
                    "cfo",
                    "capex",
                    "revenue",
                ):
                    if col in hist_view.columns:
                        numeric_cols[col] = pd.to_numeric(hist_view[col], errors="coerce")
                    else:
                        numeric_cols[col] = pd.Series([float("nan")] * len(hist_view), dtype="float64")

                def _series_map(ser: pd.Series) -> Dict[pd.Timestamp, Any]:
                    return {
                        quarter_list[idx]: (float(val) if pd.notna(val) else None)
                        for idx, val in enumerate(ser.tolist())
                    }

                def _rolling_sum_map(ser: pd.Series) -> Dict[pd.Timestamp, Any]:
                    summed = ser.rolling(4, min_periods=4).sum()
                    counts = ser.rolling(4, min_periods=4).count()
                    return {
                        quarter_list[idx]: (
                            float(summed.iloc[idx])
                            if idx < len(quarter_list) and pd.notna(summed.iloc[idx]) and counts.iloc[idx] == 4
                            else None
                        )
                        for idx in range(len(quarter_list))
                    }

                q_latest = quarter_list[-1]
                latest_row = hist_view.iloc[-1]
                last4 = hist_view.iloc[max(0, len(hist_view) - 4) :].copy()
                latest_context = {
                    "quarter": q_latest,
                    "row": latest_row,
                    "last4": last4,
                    "last4_quarters": [pd.Timestamp(q) for q in last4["quarter"].tolist()],
                }
                last4_context = {
                    "frame": last4,
                    "quarters": list(latest_context["last4_quarters"]),
                }
                core_maps = {
                    "quarters": quarter_list,
                    "cash": _series_map(numeric_cols["cash"]),
                    "debt_core": _series_map(numeric_cols["debt_core"]),
                    "ebitda_ttm": _rolling_sum_map(numeric_cols["ebitda"]),
                    "interest_paid_ttm": _rolling_sum_map(numeric_cols["interest_paid"]),
                    "interest_expense_net_ttm": _rolling_sum_map(numeric_cols["interest_expense_net"]),
                    "cfo_ttm": _rolling_sum_map(numeric_cols["cfo"]),
                    "capex_ttm": _rolling_sum_map(numeric_cols["capex"]),
                    "revenue_ttm": _rolling_sum_map(numeric_cols["revenue"]),
                }

        rev_commit_q: Dict[pd.Timestamp, Any] = {}
        rev_drawn_q: Dict[pd.Timestamp, Any] = {}
        rev_avail_q: Dict[pd.Timestamp, Any] = {}
        rev_lc_q: Dict[pd.Timestamp, Any] = {}
        rev_facility_q: Dict[pd.Timestamp, Any] = {}
        rev_keys: List[pd.Timestamp] = []
        if revolver_history is not None and not revolver_history.empty:
            sub = revolver_history.copy()
            sub["quarter"] = pd.to_datetime(sub["quarter"], errors="coerce")
            rev_commit_q = pd.to_numeric(sub.get("revolver_commitment"), errors="coerce").groupby(sub["quarter"]).max().to_dict()
            rev_drawn_q = pd.to_numeric(sub.get("revolver_drawn"), errors="coerce").groupby(sub["quarter"]).max().to_dict()
            rev_avail_q = pd.to_numeric(sub.get("revolver_availability"), errors="coerce").groupby(sub["quarter"]).max().to_dict()
            rev_lc_q = (
                pd.to_numeric(sub.get("revolver_letters_of_credit"), errors="coerce").groupby(sub["quarter"]).max().to_dict()
                if "revolver_letters_of_credit" in sub.columns
                else {}
            )
            rev_facility_q = (
                pd.to_numeric(sub.get("revolver_facility_size"), errors="coerce").groupby(sub["quarter"]).max().to_dict()
                if "revolver_facility_size" in sub.columns
                else {}
            )
        elif revolver_df is not None and not revolver_df.empty:
            sub = revolver_df.copy()
            sub["quarter"] = pd.to_datetime(sub["quarter"], errors="coerce")
            sub["filed"] = pd.to_datetime(sub.get("filed"), errors="coerce")
            form_pri = {"10-Q": 0, "10-Q/A": 1, "10-K": 0, "10-K/A": 1, "8-K": 2, "8-K/A": 3}
            sub["form_pri"] = sub.get("form").map(form_pri).fillna(5)
            sub["commitment"] = pd.to_numeric(sub.get("revolver_commitment"), errors="coerce")
            sub["has_commit"] = sub["commitment"].notna()
            sub["has_avail"] = sub.get("revolver_availability").notna()
            sub = sub.dropna(subset=["quarter"]).sort_values(
                ["quarter", "has_commit", "commitment", "has_avail", "form_pri", "filed"],
                ascending=[True, False, False, False, True, False],
            )
            rev_commit_q = sub.groupby("quarter")["commitment"].max().to_dict()
            rev_drawn_q = sub.groupby("quarter")["revolver_drawn"].max().to_dict()
            rev_avail_q = sub.groupby("quarter")["revolver_availability"].max().to_dict()
            rev_lc_q = sub.groupby("quarter")["revolver_lc"].max().to_dict() if "revolver_lc" in sub.columns else {}
            rev_facility_q = sub.groupby("quarter")["revolver_facility_size"].max().to_dict() if "revolver_facility_size" in sub.columns else {}
        if rev_commit_q:
            rev_keys = sorted(pd.Timestamp(k) for k in rev_commit_q.keys())

        adj_view = pd.DataFrame()
        adj_ebit_q: Dict[pd.Timestamp, Any] = {}
        adj_ebit_ttm_q: Dict[pd.Timestamp, Any] = {}
        adj_ebitda_q: Dict[pd.Timestamp, Any] = {}
        adj_ebitda_ttm_q: Dict[pd.Timestamp, Any] = {}
        if adj_metrics is not None and not adj_metrics.empty and "quarter" in adj_metrics.columns:
            adj_view = adj_metrics.copy()
            adj_view["quarter"] = pd.to_datetime(adj_view["quarter"], errors="coerce")
            adj_view = adj_view[adj_view["quarter"].notna()].sort_values("quarter").reset_index(drop=True)
            if "adj_ebit" in adj_view.columns:
                adj_view["adj_ebit_num"] = pd.to_numeric(adj_view["adj_ebit"], errors="coerce")
                adj_ebit_series = (
                    adj_view.drop_duplicates(subset=["quarter"], keep="last")
                    .set_index("quarter")["adj_ebit_num"]
                )
                adj_ebit_q = {
                    pd.Timestamp(q): (float(v) if pd.notna(v) else None)
                    for q, v in adj_ebit_series.items()
                }
            if "adj_ebitda" in adj_view.columns:
                adj_view["adj_ebitda_num"] = pd.to_numeric(adj_view["adj_ebitda"], errors="coerce")
                adj_q = (
                    adj_view.drop_duplicates(subset=["quarter"], keep="last")
                    .set_index("quarter")["adj_ebitda_num"]
                )
                adj_ebitda_q = {
                    pd.Timestamp(q): (float(v) if pd.notna(v) else None)
                    for q, v in adj_q.items()
                }
        if core_maps.get("quarters"):
            quarter_list = list(core_maps.get("quarters") or [])
            if adj_ebit_q:
                aligned_adj_ebit = pd.Series(
                    [adj_ebit_q.get(q) for q in quarter_list],
                    index=quarter_list,
                    dtype="float64",
                )
                adj_ebit_sum = aligned_adj_ebit.rolling(4, min_periods=4).sum()
                adj_ebit_count = aligned_adj_ebit.rolling(4, min_periods=4).count()
                adj_ebit_ttm_q = {
                    quarter_list[idx]: (
                        float(adj_ebit_sum.iloc[idx])
                        if pd.notna(adj_ebit_sum.iloc[idx]) and adj_ebit_count.iloc[idx] == 4
                        else None
                    )
                    for idx in range(len(quarter_list))
                }
            if adj_ebitda_q:
                aligned_adj = pd.Series(
                    [adj_ebitda_q.get(q) for q in quarter_list],
                    index=quarter_list,
                    dtype="float64",
                )
                adj_sum = aligned_adj.rolling(4, min_periods=4).sum()
                adj_count = aligned_adj.rolling(4, min_periods=4).count()
                adj_ebitda_ttm_q = {
                    quarter_list[idx]: (
                        float(adj_sum.iloc[idx])
                        if pd.notna(adj_sum.iloc[idx]) and adj_count.iloc[idx] == 4
                        else None
                    )
                    for idx in range(len(quarter_list))
                }

        derived.valuation_hist_view = hist_view
        derived.valuation_hist_indexed = hist_indexed
        derived.valuation_latest_context = latest_context
        derived.valuation_last4_context = last4_context
        derived.valuation_core_maps = core_maps
        derived.valuation_adj_metrics_view = adj_view
        derived.valuation_revolver_maps = {
            "commitment": rev_commit_q,
            "drawn": rev_drawn_q,
            "availability": rev_avail_q,
            "letters_of_credit": rev_lc_q,
            "facility_size": rev_facility_q,
            "keys": rev_keys,
        }
        derived.valuation_adj_ebit_q = adj_ebit_q
        derived.valuation_adj_ebit_ttm_q = adj_ebit_ttm_q
        derived.valuation_adj_ebitda_q = adj_ebitda_q
        derived.valuation_adj_ebitda_ttm_q = adj_ebitda_ttm_q


def ensure_valuation_inputs(ctx: WriterContext) -> None:
    callbacks = ctx.callbacks
    state = ctx.state
    if ctx.derived.leverage_df is not None:
        return
    hist = ctx.inputs.hist
    adj_metrics = ctx.inputs.adj_metrics
    with timed_writer_stage(
        ctx.writer_timings,
        "write_excel.derive.valuation_inputs",
        enabled=bool(ctx.inputs.profile_timings),
    ):
        _ensure_valuation_source_views(ctx)
        lev_rows: List[Dict[str, Any]] = []
        with timed_writer_stage(
            ctx.writer_timings,
            "write_excel.derive.valuation_inputs.net_leverage_text_map",
            enabled=bool(ctx.inputs.profile_timings),
        ):
            net_lev_text_map = callbacks.extract_adj_net_leverage_text_map()
        h = ctx.derived.valuation_hist_view if ctx.derived.valuation_hist_view is not None else pd.DataFrame()
        with timed_writer_stage(
            ctx.writer_timings,
            "write_excel.derive.valuation_inputs.leverage_frame",
            enabled=bool(ctx.inputs.profile_timings),
        ):
            if not h.empty:
                valuation_core_maps = ctx.derived.valuation_core_maps or {}
                quarter_list = list(valuation_core_maps.get("quarters") or [])
                cash_q = dict(valuation_core_maps.get("cash") or {})
                debt_core_q = dict(valuation_core_maps.get("debt_core") or {})
                ebitda_ttm_q = dict(valuation_core_maps.get("ebitda_ttm") or {})
                interest_cash_ttm_q = dict(valuation_core_maps.get("interest_paid_ttm") or {})
                interest_pnl_ttm_q = dict(valuation_core_maps.get("interest_expense_net_ttm") or {})
                revolver_maps = ctx.derived.valuation_revolver_maps or {}
                rev_commit_q = revolver_maps.get("commitment", {})
                rev_drawn_q = revolver_maps.get("drawn", {})
                rev_avail_q = revolver_maps.get("availability", {})
                rev_lc_q = revolver_maps.get("letters_of_credit", {})
                rev_facility_q = revolver_maps.get("facility_size", {})
                rev_keys = list(revolver_maps.get("keys", []))

                def _carry_forward(
                    q: pd.Timestamp,
                ) -> Tuple[Optional[float], Optional[float], Optional[float], Optional[float], str]:
                    note = ""
                    if not rev_keys:
                        return None, None, None, None, note
                    idx = bisect_right(rev_keys, q) - 1
                    if idx < 0:
                        return None, None, None, None, note
                    last_q = rev_keys[idx]
                    if (q - last_q).days > 1860:
                        return None, None, None, None, note
                    note = "revolver carry-forward"
                    return (rev_commit_q.get(last_q), None, None, None, note)

                adj_ebitda_ttm_q = ctx.derived.valuation_adj_ebitda_ttm_q or {}

                for idx in range(len(quarter_list)):
                    if idx < 3:
                        continue
                    q = quarter_list[idx]
                    ebitda_ttm = ebitda_ttm_q.get(q)
                    interest_cash_ttm = interest_cash_ttm_q.get(q)
                    interest_pnl_ttm = interest_pnl_ttm_q.get(q)
                    cash = cash_q.get(q)
                    debt_core = debt_core_q.get(q)
                    net_debt = (debt_core - cash) if pd.notna(debt_core) and pd.notna(cash) else None
                    ebitda_ok = ebitda_ttm is not None and ebitda_ttm >= 50_000_000 and ebitda_ttm != 0
                    adj_ebitda_ttm = adj_ebitda_ttm_q.get(q) if adj_ebitda_ttm_q else None
                    adj_ok = adj_ebitda_ttm is not None and adj_ebitda_ttm >= 50_000_000 and adj_ebitda_ttm != 0
                    net_lev_gaap = (net_debt / ebitda_ttm) if (net_debt is not None and ebitda_ok) else None
                    net_lev_adj = (net_debt / adj_ebitda_ttm) if (net_debt is not None and adj_ok) else None
                    q_key = pd.Timestamp(q).to_period("Q").end_time.normalize()
                    net_lev_text = net_lev_text_map.get(q_key)
                    net_lev = net_lev_gaap
                    cov_pnl = None
                    pnl_note = ""
                    coverage_ebitda = ebitda_ttm if ebitda_ok else None
                    if coverage_ebitda is None:
                        pnl_note = "gaap ebitda_ttm < $50m or unavailable; leverage/coverage N/A"
                    elif interest_pnl_ttm is not None and interest_pnl_ttm <= 0:
                        pnl_note = "pnl interest <= 0; coverage not meaningful"
                    else:
                        cov_pnl = (
                            (coverage_ebitda / abs(interest_pnl_ttm))
                            if (coverage_ebitda is not None and interest_pnl_ttm not in (None, 0))
                            else None
                        )
                    cov_cash = (
                        (coverage_ebitda / abs(interest_cash_ttm))
                        if (coverage_ebitda is not None and interest_cash_ttm not in (None, 0))
                        else None
                    )
                    rev_note = ""
                    rev_commit = rev_commit_q.get(q)
                    rev_facility = rev_facility_q.get(q)
                    rev_drawn = rev_drawn_q.get(q)
                    rev_avail = rev_avail_q.get(q)
                    rev_lc = rev_lc_q.get(q)
                    if rev_commit is None and rev_drawn is None and rev_avail is None:
                        rev_commit, rev_drawn, rev_avail, rev_lc, rev_note = _carry_forward(q)
                    if rev_avail is None and rev_commit is not None and rev_drawn is not None and rev_lc is not None:
                        rev_avail = float(rev_commit) - float(rev_drawn) - float(rev_lc)
                    elif rev_avail is None and rev_commit is not None and rev_drawn is not None and rev_lc is None:
                        rev_avail = float(rev_commit) - float(rev_drawn)
                        rev_note = (rev_note + "; " if rev_note else "") + "lc_missing"
                    liquidity = None
                    liquidity_note = ""
                    if pd.notna(cash) and pd.notna(rev_avail):
                        liquidity = float(cash) + float(rev_avail)
                    elif pd.notna(cash) and pd.isna(rev_avail):
                        liquidity = float(cash)
                        liquidity_note = "revolver availability missing; liquidity = cash only"
                    lev_rows.append(
                        {
                            "quarter": q.date(),
                            "corporate_net_debt": net_debt,
                            "ebitda_ttm": ebitda_ttm,
                            "corporate_net_leverage": net_lev,
                            "corporate_net_leverage_calc": net_lev_adj,
                            "corporate_net_leverage_text": net_lev_text,
                            "corporate_net_leverage_basis": "gaap_ebitda_ttm" if net_lev_gaap is not None else "missing_gaap_ebitda_ttm",
                            "adj_ebitda_ttm": adj_ebitda_ttm,
                            "corporate_net_leverage_adj": net_lev_adj,
                            "corporate_net_leverage_adj_basis": "adjusted_ebitda_ttm" if net_lev_adj is not None else "missing_adjusted_ebitda_ttm",
                            "interest_expense_net_ttm": interest_pnl_ttm,
                            "interest_coverage_pnl": cov_pnl,
                            "interest_coverage_pnl_basis": (
                                "gaap_ebitda_ttm/pnl_interest_ttm"
                                if cov_pnl is not None
                                else ("missing_gaap_ebitda_ttm" if coverage_ebitda is None else "missing_or_nonmeaningful_pnl_interest")
                            ),
                            "interest_paid_ttm": interest_cash_ttm,
                            "interest_coverage_cash": cov_cash,
                            "interest_coverage_cash_basis": (
                                "gaap_ebitda_ttm/cash_interest_ttm"
                                if cov_cash is not None
                                else ("missing_gaap_ebitda_ttm" if coverage_ebitda is None else "missing_cash_interest_ttm")
                            ),
                            "cash": cash,
                            "revolver_commitment": rev_commit,
                            "revolver_facility_size": rev_facility,
                            "revolver_drawn": rev_drawn,
                            "revolver_letters_of_credit": rev_lc,
                            "revolver_availability": rev_avail,
                            "liquidity": liquidity,
                            "note": "; ".join([n for n in [liquidity_note, pnl_note, rev_note] if n]),
                        }
                    )

        leverage_df = _sync_frame_placeholder(state.get("leverage_df"), pd.DataFrame(lev_rows))
        valuation_summary_df = pd.DataFrame()
        valuation_grid_df = pd.DataFrame()
        with timed_writer_stage(
            ctx.writer_timings,
            "write_excel.derive.valuation_inputs.valuation_frames",
            enabled=bool(ctx.inputs.profile_timings),
        ):
            try:
                latest_context = ctx.derived.valuation_latest_context or {}
                latest_row = latest_context.get("row")
                q_latest = latest_context.get("quarter")
                if latest_row is not None and q_latest is not None:
                    valuation_core_maps = ctx.derived.valuation_core_maps or {}
                    leverage_row_by_quarter = {
                        pd.Timestamp(pd.to_datetime(row.get("quarter"), errors="coerce")): row
                        for row in lev_rows
                        if pd.notna(pd.to_datetime(row.get("quarter"), errors="coerce"))
                    }
                    lev_latest = leverage_row_by_quarter.get(pd.Timestamp(q_latest)) or {}
                    ebitda_ttm = pd.to_numeric(lev_latest.get("ebitda_ttm"), errors="coerce")
                    if pd.isna(ebitda_ttm):
                        ebitda_ttm = pd.to_numeric((valuation_core_maps.get("ebitda_ttm") or {}).get(pd.Timestamp(q_latest)), errors="coerce")
                    ebitda_ttm = float(ebitda_ttm) if pd.notna(ebitda_ttm) else None

                    adj_ebitda_ttm = pd.to_numeric((ctx.derived.valuation_adj_ebitda_ttm_q or {}).get(pd.Timestamp(q_latest)), errors="coerce")
                    adj_ebitda_ttm = float(adj_ebitda_ttm) if pd.notna(adj_ebitda_ttm) else None

                    interest_paid_ttm = pd.to_numeric(lev_latest.get("interest_paid_ttm"), errors="coerce")
                    if pd.isna(interest_paid_ttm):
                        interest_paid_ttm = pd.to_numeric((valuation_core_maps.get("interest_paid_ttm") or {}).get(pd.Timestamp(q_latest)), errors="coerce")
                    interest_paid_ttm = float(interest_paid_ttm) if pd.notna(interest_paid_ttm) else None

                    debt_core = pd.to_numeric(latest_row.get("debt_core"), errors="coerce")
                    cash = pd.to_numeric(latest_row.get("cash"), errors="coerce")
                    net_debt = (float(debt_core) - float(cash)) if pd.notna(debt_core) and pd.notna(cash) else None
                    shares_out = pd.to_numeric(latest_row.get("shares_outstanding"), errors="coerce")
                    shares_dil = pd.to_numeric(latest_row.get("shares_diluted"), errors="coerce")
                    if pd.isna(shares_dil):
                        shares_dil = shares_out

                    cfo_ttm = pd.to_numeric((valuation_core_maps.get("cfo_ttm") or {}).get(pd.Timestamp(q_latest)), errors="coerce")
                    capex_ttm = pd.to_numeric((valuation_core_maps.get("capex_ttm") or {}).get(pd.Timestamp(q_latest)), errors="coerce")
                    revenue_ttm = pd.to_numeric((valuation_core_maps.get("revenue_ttm") or {}).get(pd.Timestamp(q_latest)), errors="coerce")

                    hist_latest = {
                        "quarter": pd.Timestamp(q_latest).date(),
                        "shares_outstanding_m": (float(shares_out) / 1e6) if pd.notna(shares_out) else None,
                        "shares_diluted_m": (float(shares_dil) / 1e6) if pd.notna(shares_dil) else None,
                        "debt_core_m": (float(debt_core) / 1e6) if pd.notna(debt_core) else None,
                        "cash_m": (float(cash) / 1e6) if pd.notna(cash) else None,
                        "net_debt_m": (net_debt / 1e6) if net_debt is not None else None,
                        "ebitda_ttm_m": (ebitda_ttm / 1e6) if ebitda_ttm is not None else None,
                        "adj_ebitda_ttm_m": (adj_ebitda_ttm / 1e6) if adj_ebitda_ttm is not None else None,
                        "fcf_ttm_m": ((float(cfo_ttm) - float(capex_ttm)) / 1e6) if pd.notna(cfo_ttm) and pd.notna(capex_ttm) else None,
                        "interest_paid_ttm_m": (interest_paid_ttm / 1e6) if interest_paid_ttm is not None else None,
                        "revenue_ttm_m": (float(revenue_ttm) / 1e6) if pd.notna(revenue_ttm) else None,
                        "capex_ttm_m": (float(capex_ttm) / 1e6) if pd.notna(capex_ttm) else None,
                    }
                    scenario_inputs = {
                        "target_ev_ebitda": 6.0,
                        "target_ev_yield": 0.10,
                        "maint_capex_ratio": 0.70,
                        "recurring_cash_costs_m": 0.0,
                        "wc_normalization_m": 0.0,
                        "scenarios": {
                            "base": {"rev_growth": 0.00, "margin_delta": 0.00, "refi_norm_m": 0.0, "buyback_m": 0.0, "ev_multiple": 6.0},
                            "bull": {"rev_growth": 0.02, "margin_delta": 0.01, "refi_norm_m": 15.0, "buyback_m": 2.0, "ev_multiple": 7.0},
                            "bear": {"rev_growth": -0.03, "margin_delta": -0.01, "refi_norm_m": -10.0, "buyback_m": -1.0, "ev_multiple": 5.0},
                        },
                    }
                    vout = valuation_engine(price=state["price"], scenario_inputs=scenario_inputs, hist_latest=hist_latest)
                    valuation_summary_df, valuation_grid_df = valuation_to_frames(vout)
            except Exception as exc:
                valuation_summary_df = pd.DataFrame([{"metric": "engine_error", "value": str(exc), "unit": "", "section": "qa"}])
                valuation_grid_df = pd.DataFrame()

        valuation_summary_df = _sync_frame_placeholder(state.get("valuation_summary_df"), valuation_summary_df)
        valuation_grid_df = _sync_frame_placeholder(state.get("valuation_grid_df"), valuation_grid_df)
        state["leverage_df"] = leverage_df
        state["valuation_summary_df"] = valuation_summary_df
        state["valuation_grid_df"] = valuation_grid_df
        ctx.derived.leverage_df = leverage_df
        ctx.derived.valuation_summary_df = valuation_summary_df
        ctx.derived.valuation_grid_df = valuation_grid_df


def ensure_hidden_value_inputs(ctx: WriterContext) -> None:
    data = ctx.data
    callbacks = ctx.callbacks
    state = ctx.state
    if ctx.derived.flags_df is not None:
        return
    ensure_valuation_inputs(ctx)
    with timed_writer_stage(
        ctx.writer_timings,
        "write_excel.derive.hidden_value_inputs",
        enabled=bool(ctx.inputs.profile_timings),
    ):
        leverage_df = ctx.derived.leverage_df if ctx.derived.leverage_df is not None else state.get("leverage_df")
        signals_base_df = build_signals_base(
            hist=ctx.inputs.hist,
            adj_metrics=ctx.inputs.adj_metrics,
            leverage_df=leverage_df,
            debt_tranches=ctx.inputs.debt_tranches,
            price=ctx.inputs.price,
        )
        flags_df, flags_audit_df, flags_recompute_df = build_hidden_value_outputs(
            hist=ctx.inputs.hist,
            adj_metrics=ctx.inputs.adj_metrics,
            leverage_df=leverage_df,
            debt_tranches=ctx.inputs.debt_tranches,
            signals_base=signals_base_df,
            price=ctx.inputs.price,
            max_flags=10,
        )

        def _ordered_hidden_value_flags_frame(
            current_flags: Optional[pd.DataFrame],
            audit_flags: Optional[pd.DataFrame],
        ) -> pd.DataFrame:
            def _metric_piece_for_visible_support(metric_key: str, value_in: Any) -> str:
                try:
                    fval = float(value_in)
                except Exception:
                    return ""
                if metric_key in {"ebit_growth_yoy", "ebitda_growth_yoy", "shares_yoy", "pos_fcf_ratio", "fcf_yield"}:
                    sign = "+" if fval > 0 else ""
                    label_map = {
                        "ebit_growth_yoy": "EBIT YoY",
                        "ebitda_growth_yoy": "EBITDA YoY",
                        "shares_yoy": "Shares YoY",
                        "pos_fcf_ratio": "Positive FCF ratio",
                        "fcf_yield": "FCF yield",
                    }
                    return f"{label_map.get(metric_key, metric_key)} {sign}{fval * 100.0:.1f}%"
                if metric_key == "debt_drop_pct":
                    return f"Net debt drop {fval * 100.0:.1f}%"
                if metric_key in {"adj_margin_ttm"}:
                    return f"Adj margin TTM {fval * 100.0:.1f}%"
                if metric_key == "margin_yoy_bps":
                    sign = "+" if fval > 0 else ""
                    return f"Margin YoY {sign}{fval:,.0f}bps"
                if metric_key == "margin_streak":
                    return f"Margin streak {fval:.0f} quarters"
                if metric_key == "interest_coverage":
                    return f"Interest cover {fval:.2f}x"
                if metric_key == "leverage_ratio":
                    return f"Leverage {fval:.2f}x"
                if metric_key == "corporate_net_debt":
                    return f"Net debt ${fval / 1_000_000.0:,.1f}m"
                if metric_key in {"ebitda_ttm", "ebit_ttm"}:
                    label = "EBITDA TTM" if metric_key == "ebitda_ttm" else "EBIT TTM"
                    return f"{label} ${fval / 1_000_000.0:,.1f}m"
                if metric_key == "fcf_ttm_pos_years":
                    return f"Positive FCF years {fval:.0f}"
                return ""

            def _visible_support_from_metrics(metrics_json_in: Any) -> str:
                raw_txt = str(metrics_json_in or "").strip()
                if not raw_txt.startswith("{"):
                    return ""
                try:
                    metrics_obj = json.loads(raw_txt)
                except Exception:
                    metrics_obj = {}
                if not isinstance(metrics_obj, dict):
                    return ""
                metric_order = [
                    "ebit_growth_yoy",
                    "ebitda_growth_yoy",
                    "shares_yoy",
                    "adj_margin_ttm",
                    "margin_yoy_bps",
                    "margin_streak",
                    "fcf_ttm_pos_years",
                    "pos_fcf_ratio",
                    "fcf_yield",
                    "interest_coverage",
                    "debt_drop_pct",
                    "leverage_ratio",
                    "corporate_net_debt",
                    "ebitda_ttm",
                    "ebit_ttm",
                ]
                parts: List[str] = []
                for metric_key in metric_order:
                    metric_val = metrics_obj.get(metric_key)
                    if metric_val in (None, "", "null"):
                        continue
                    piece = _metric_piece_for_visible_support(metric_key, metric_val)
                    if piece:
                        parts.append(piece)
                    if len(parts) >= 3:
                        break
                if not parts:
                    return ""
                return ", ".join(parts)

            active_df = current_flags.copy() if isinstance(current_flags, pd.DataFrame) and not current_flags.empty else pd.DataFrame()
            audit_df = audit_flags.copy() if isinstance(audit_flags, pd.DataFrame) and not audit_flags.empty else pd.DataFrame()
            if audit_df.empty:
                return active_df
            if "quarter" in audit_df.columns:
                audit_df["quarter"] = pd.to_datetime(audit_df["quarter"], errors="coerce")
                latest_q = audit_df["quarter"].dropna().max()
                if pd.notna(latest_q):
                    audit_df = audit_df[audit_df["quarter"] == latest_q].copy()
            preferred_order = {
                "A": 1,
                "C": 2,
                "E": 3,
                "F": 4,
                "B": 5,
                "D": 6,
                "G": 7,
            }
            active_by_code: Dict[str, Dict[str, Any]] = {}
            if not active_df.empty and "flag_code" in active_df.columns:
                for _, rec in active_df.iterrows():
                    code = str(rec.get("flag_code") or "").strip().upper()
                    if code:
                        active_by_code[code] = rec.to_dict()
            rows: List[Dict[str, Any]] = []
            for _, rec in audit_df.iterrows():
                code = str(rec.get("flag_id") or "").strip().upper()
                if not code:
                    continue
                active = active_by_code.get(code, {})
                title = str(
                    active.get("title")
                    or active.get("Title")
                    or rec.get("flag_name")
                    or f"Flag {code}"
                ).strip()
                quarter_val = rec.get("quarter")
                quarter_txt = ""
                if quarter_val not in (None, ""):
                    q_ts = pd.to_datetime(quarter_val, errors="coerce")
                    quarter_txt = q_ts.strftime("%Y-%m-%d") if pd.notna(q_ts) else str(quarter_val)
                rows.append(
                    {
                        "rank": preferred_order.get(code, 90 + len(rows)),
                        "flag_code": code,
                        "title": title,
                        "score": active.get("score", 0),
                        "severity": active.get("severity") or "Info",
                        "as_of_quarter": active.get("as_of_quarter") or quarter_txt,
                        "evidence_1": active.get("evidence_1") or "",
                        "evidence_2": active.get("evidence_2") or "",
                        "evidence_3": active.get("evidence_3") or "",
                        "metrics_json": active.get("metrics_json") or rec.get("inputs_json") or "",
                        "visible_support": _visible_support_from_metrics(active.get("metrics_json") or rec.get("inputs_json") or ""),
                    }
                )
            if not rows:
                return active_df
            ordered_df = pd.DataFrame(rows).sort_values(["rank", "flag_code"], ascending=[True, True], na_position="last").reset_index(drop=True)
            ordered_df["rank"] = range(1, len(ordered_df) + 1)
            return ordered_df

        flags_df = _ordered_hidden_value_flags_frame(flags_df, flags_audit_df)
        if flags_df is None or flags_df.empty:
            flags_df = callbacks.build_hidden_value_flags_fallback(flags_audit_df)

        qa_checks = data.qa_checks if isinstance(data.qa_checks, pd.DataFrame) else state.get("qa_checks")
        if qa_checks is None or qa_checks.empty:
            qa_checks = pd.DataFrame()
        if flags_audit_df is not None and not flags_audit_df.empty:
            bad_audit = flags_audit_df[flags_audit_df["qa_severity"].astype(str).str.lower().isin(["warn", "fail"])].copy()
            if not bad_audit.empty:
                qa_rows = []
                for _, r in bad_audit.iterrows():
                    qa_rows.append(
                        {
                            "quarter": r.get("quarter"),
                            "metric": f"hidden_flag_{r.get('flag_id')}",
                            "check": "hidden_flag_audit",
                            "status": str(r.get("qa_severity")).lower(),
                            "message": str(r.get("qa_message") or ""),
                        }
                    )
                if qa_rows:
                    qa_checks = pd.concat([qa_checks, pd.DataFrame(qa_rows)], ignore_index=True)
        if flags_recompute_df is not None and not flags_recompute_df.empty:
            bad_rec = flags_recompute_df[flags_recompute_df["qa_severity"].astype(str).str.lower() == "fail"].copy()
            if not bad_rec.empty:
                qa_rows = []
                for _, r in bad_rec.iterrows():
                    qa_rows.append(
                        {
                            "quarter": r.get("quarter"),
                            "metric": f"hidden_flag_{r.get('flag_id')}",
                            "check": "hidden_flag_recompute",
                            "status": "fail",
                            "message": str(r.get("qa_message") or "Hidden flag recompute mismatch"),
                        }
                    )
                qa_checks = pd.concat([qa_checks, pd.DataFrame(qa_rows)], ignore_index=True)

        signals_base_df = _sync_frame_placeholder(state.get("signals_base_df"), signals_base_df)
        flags_df = _sync_frame_placeholder(state.get("flags_df"), flags_df)
        flags_audit_df = _sync_frame_placeholder(state.get("flags_audit_df"), flags_audit_df)
        flags_recompute_df = _sync_frame_placeholder(state.get("flags_recompute_df"), flags_recompute_df)
        state["signals_base_df"] = signals_base_df
        state["flags_df"] = flags_df
        state["flags_audit_df"] = flags_audit_df
        state["flags_recompute_df"] = flags_recompute_df
        state["qa_checks"] = qa_checks
        data.qa_checks = qa_checks
        ctx.derived.signals_base_df = signals_base_df
        ctx.derived.flags_df = flags_df
        ctx.derived.flags_audit_df = flags_audit_df
        ctx.derived.flags_recompute_df = flags_recompute_df


def ensure_summary_inputs(ctx: WriterContext) -> None:
    callbacks = ctx.callbacks
    state = ctx.state
    if ctx.derived.summary_df is not None:
        return
    ensure_valuation_inputs(ctx)
    with timed_writer_stage(
        ctx.writer_timings,
        "write_excel.derive.summary_inputs",
        enabled=bool(ctx.inputs.profile_timings),
    ):
        summary_df = _sync_frame_placeholder(state.get("summary_df"), callbacks.build_summary())
        state["summary_df"] = summary_df
        ctx.derived.summary_df = summary_df


def ensure_raw_data_inputs(ctx: WriterContext) -> None:
    callbacks = ctx.callbacks
    state = ctx.state
    if ctx.derived.facts_long is not None:
        return
    with timed_writer_stage(
        ctx.writer_timings,
        "write_excel.derive.raw_data_inputs",
        enabled=bool(ctx.inputs.profile_timings),
    ):
        facts_long = _sync_frame_placeholder(state.get("facts_long"), callbacks.build_facts_long())
        lineitem_map = _sync_frame_placeholder(state.get("lineitem_map"), callbacks.build_lineitem_map())
        period_index = _sync_frame_placeholder(state.get("period_index"), callbacks.build_period_index(12))
        ng_bridge = _sync_frame_placeholder(state.get("ng_bridge"), callbacks.build_ng_bridge(ctx.inputs.adj_metrics, ctx.inputs.adj_breakdown))
        relaxed_df = (
            callbacks.build_ng_bridge(ctx.inputs.adj_metrics_relaxed, ctx.inputs.adj_breakdown_relaxed)
            if ctx.inputs.excel_mode == "full"
            else pd.DataFrame()
        )
        ng_bridge_relaxed = _sync_frame_placeholder(state.get("ng_bridge_relaxed"), relaxed_df)
        state["facts_long"] = facts_long
        state["lineitem_map"] = lineitem_map
        state["period_index"] = period_index
        state["ng_bridge"] = ng_bridge
        state["ng_bridge_relaxed"] = ng_bridge_relaxed
        ctx.derived.facts_long = facts_long
        ctx.derived.lineitem_map = lineitem_map
        ctx.derived.period_index = period_index
        ctx.derived.ng_bridge = ng_bridge
        ctx.derived.ng_bridge_relaxed = ng_bridge_relaxed


def ensure_ui_evidence(ctx: WriterContext) -> None:
    callbacks = ctx.callbacks
    state = ctx.state
    if ctx.derived.quarter_notes_evidence_df is not None:
        return
    with timed_writer_stage(
        ctx.writer_timings,
        "write_excel.derive.ui_evidence",
        enabled=bool(ctx.inputs.profile_timings),
    ):
        quarter_notes_evidence_df = _sync_frame_placeholder(
            state.get("quarter_notes_evidence_df"),
            callbacks.build_qn_evidence_src(),
        )
        promise_evidence_df = _sync_frame_placeholder(
            state.get("promise_evidence_df"),
            callbacks.build_promise_evidence_src(),
        )
        state["quarter_notes_evidence_df"] = quarter_notes_evidence_df
        state["promise_evidence_df"] = promise_evidence_df
        ctx.derived.quarter_notes_evidence_df = quarter_notes_evidence_df
        ctx.derived.promise_evidence_df = promise_evidence_df


def write_raw_data_sheets(ctx: WriterContext) -> None:
    data = ctx.data
    callbacks = ctx.callbacks
    ensure_driver_inputs(ctx)
    ensure_raw_data_inputs(ctx)

    callbacks.write_sheet("History_Q", ctx.inputs.hist)
    if data.enable_operating_drivers_sheet:
        callbacks.write_operating_drivers_raw_sheet(data.operating_driver_history_rows)
    if data.enable_economics_market_raw_sheet:
        callbacks.write_economics_market_raw_sheet(data.economics_market_rows)
    callbacks.write_sheet("Adjusted_Metrics", ctx.inputs.adj_metrics)
    callbacks.write_sheet("Adjustments_Breakdown", ctx.inputs.adj_breakdown)
    callbacks.write_sheet("NonGAAP_Files", ctx.inputs.non_gaap_files)
    callbacks.write_sheet("Slides_Segments", ctx.inputs.slides_segments)
    callbacks.write_sheet("Slides_Debt_Profile", ctx.inputs.slides_debt)
    callbacks.write_sheet("NonGAAP_Bridge", ctx.require_derived_frame("ng_bridge"))
    if ctx.inputs.excel_mode == "full" and not ctx.inputs.adj_metrics_relaxed.empty:
        callbacks.write_sheet("Adjusted_Metrics_Relaxed", ctx.inputs.adj_metrics_relaxed)
        callbacks.write_sheet("Adjustments_Breakdown_Relaxed", ctx.inputs.adj_breakdown_relaxed)
        callbacks.write_sheet("NonGAAP_Files_Relaxed", ctx.inputs.non_gaap_files_relaxed)
        callbacks.write_sheet(
            "NonGAAP_Bridge_Relaxed",
            ctx.require_derived_frame("ng_bridge_relaxed"),
        )

    callbacks.write_sheet("DATA_Facts_Long", ctx.require_derived_frame("facts_long"))
    callbacks.write_sheet("DATA_LineItem_Map", ctx.require_derived_frame("lineitem_map"))
    callbacks.write_sheet("DATA_Period_Index", ctx.require_derived_frame("period_index"))
    if not data.data_is_rules_df.empty:
        callbacks.write_sheet("DATA_IS_Rules", data.data_is_rules_df)


def write_qa_sheets(ctx: WriterContext, ui_qa_rows: List[Dict[str, Any]]) -> None:
    data = ctx.data
    callbacks = ctx.callbacks
    state = ctx.state
    info_log = data.info_log
    qa_checks = data.qa_checks
    needs_review = ctx.inputs.needs_review

    if ctx.ui_info_rows:
        ui_info_df = pd.DataFrame(ctx.ui_info_rows)
        for col, default in [
            ("quarter", pd.NaT),
            ("metric", ""),
            ("severity", "info"),
            ("message", ""),
            ("source", ""),
        ]:
            if col not in ui_info_df.columns:
                ui_info_df[col] = default
        ui_info_df = ui_info_df[["quarter", "metric", "severity", "message", "source"]].copy()
        if info_log is None or info_log.empty:
            info_log = ui_info_df
        else:
            info_log = pd.concat([info_log, ui_info_df], ignore_index=True)

    try:
        qa_info_rows = callbacks.run_latest_quarter_qa()
    except Exception as qa_ex:
        qa_info_rows = [
            {
                "quarter": pd.NaT,
                "metric": "QA_QTR",
                "severity": "warn",
                "message": f"Latest-quarter QA failed: {qa_ex}",
                "source": "pipeline",
            }
        ]
    if qa_info_rows:
        qa_info_df = pd.DataFrame(qa_info_rows)
        for col, default in [
            ("quarter", pd.NaT),
            ("metric", ""),
            ("severity", "info"),
            ("message", ""),
            ("source", ""),
        ]:
            if col not in qa_info_df.columns:
                qa_info_df[col] = default
        qa_info_df = qa_info_df[["quarter", "metric", "severity", "message", "source"]].copy()
        if info_log is None or info_log.empty:
            info_log = qa_info_df
        else:
            info_log = pd.concat([info_log, qa_info_df], ignore_index=True)
        if qa_checks is None or qa_checks.empty:
            qa_checks = qa_info_df.copy()
        else:
            qa_checks = pd.concat([qa_checks, qa_info_df], ignore_index=True)
        qa_needs_review = qa_info_df[qa_info_df["severity"].astype(str).str.lower().isin(["warn", "fail"])].copy()
        if not qa_needs_review.empty:
            if needs_review is None or needs_review.empty:
                needs_review = qa_needs_review.copy()
            else:
                for col in needs_review.columns:
                    if col not in qa_needs_review.columns:
                        qa_needs_review[col] = pd.NA
                for col in qa_needs_review.columns:
                    if col not in needs_review.columns:
                        needs_review[col] = pd.NA
                qa_needs_review = qa_needs_review[needs_review.columns].copy()
                needs_review = pd.concat([needs_review, qa_needs_review], ignore_index=True)

    def _valuation_visible_quarters() -> List[pd.Timestamp]:
        out: List[pd.Timestamp] = []
        exp = dict(getattr(ctx.derived, "valuation_export_expectation", {}) or {})
        for hdr in list(exp.get("quarter_headers") or []):
            mt = re.fullmatch(r"(\d{4})-Q([1-4])", str(hdr or "").strip())
            if not mt:
                continue
            year = int(mt.group(1))
            quarter = int(mt.group(2))
            month = quarter * 3
            day = 31 if month in {3, 12} else 30
            out.append(pd.Timestamp(year=year, month=month, day=day))
        return out

    visible_quarters = _valuation_visible_quarters()
    if visible_quarters and needs_review is not None and not needs_review.empty:
        valuation_bundle = dict(getattr(ctx.derived, "valuation_precompute_bundle", {}) or {})
        capital_return_resolved = dict(valuation_bundle.get("capital_return_resolved") or {})
        valuation_audit = dict(valuation_bundle.get("valuation_audit") or {})
        nr = needs_review.copy()
        if "quarter" in nr.columns and "metric" in nr.columns:
            nr["quarter"] = pd.to_datetime(nr["quarter"], errors="coerce")
            drop_mask = pd.Series(False, index=nr.index)
            for qv in visible_quarters:
                qts = pd.Timestamp(qv).normalize()
                buyback_audit = dict((valuation_audit.get(qts) or {}).get("buyback_cash") or {})
                resolved_row = dict(capital_return_resolved.get(qts) or {})
                if not buyback_audit and not resolved_row:
                    continue
                drop_mask = drop_mask | (
                    nr["quarter"].eq(qts)
                    & nr["metric"].astype(str).str.lower().eq("buybacks_cash")
                )
            if bool(drop_mask.any()):
                nr = nr.loc[~drop_mask].copy()
        needs_review = nr

    def _export_rows_snapshot(
        df: Optional[pd.DataFrame],
        *,
        metrics: List[str],
        quarters: Optional[List[pd.Timestamp]] = None,
    ) -> List[Dict[str, Any]]:
        if df is None or df.empty or "metric" not in df.columns:
            return []
        snap = df.copy()
        snap["metric"] = snap["metric"].astype(str)
        snap = snap[snap["metric"].isin(metrics)].copy()
        if snap.empty:
            return []
        if "quarter" in snap.columns:
            snap["quarter"] = pd.to_datetime(snap["quarter"], errors="coerce")
            if quarters:
                qset = {pd.Timestamp(q).normalize() for q in quarters}
                snap = snap[snap["quarter"].isin(qset)].copy()
        keep_cols = [col for col in ["quarter", "metric", "severity", "status", "message", "source"] if col in snap.columns]
        if not keep_cols:
            return []
        snap = snap[keep_cols].copy()
        if "quarter" in snap.columns:
            snap = snap.sort_values(["quarter", "metric", "message"], kind="stable")
            snap["quarter"] = snap["quarter"].apply(lambda v: pd.Timestamp(v).strftime("%Y-%m-%d") if pd.notna(v) else "")
        return snap.to_dict("records")

    callbacks.write_sheet("SEC_Audit_Log", ctx.inputs.audit)
    callbacks.write_sheet("Info_Log", info_log)
    callbacks.write_sheet("OCR_Text_Log", ctx.inputs.ocr_log)
    callbacks.write_sheet("Needs_Review", needs_review)
    if ui_qa_rows:
        ui_qa_df = pd.DataFrame(ui_qa_rows)
        if qa_checks is None or qa_checks.empty:
            qa_checks = ui_qa_df.copy()
        else:
            qa_checks = pd.concat([qa_checks, ui_qa_df], ignore_index=True)
    callbacks.write_sheet("QA_Checks", qa_checks)
    if ctx.inputs.excel_mode == "full":
        callbacks.write_sheet("Tag_Coverage", ctx.inputs.tag_coverage)
        callbacks.write_sheet("Period_Self_Check", ctx.inputs.period_checks)
        callbacks.write_sheet("Bridge_Q", ctx.inputs.bridge_q)
        callbacks.write_sheet("Source_Manifest", ctx.inputs.manifest_df)
        callbacks.write_sheet("QFD_10K_Preview", ctx.inputs.qfd_preview)
        callbacks.write_sheet("QFD_10K_Unused", ctx.inputs.qfd_unused)

    state["info_log"] = info_log
    state["qa_checks"] = qa_checks
    ctx.inputs.needs_review = needs_review
    data.info_log = info_log
    data.qa_checks = qa_checks
    latest_visible_q = max(visible_quarters) if visible_quarters else pd.NaT
    latest_visible_q_txt = pd.Timestamp(latest_visible_q).strftime("%Y-%m-%d") if pd.notna(latest_visible_q) else ""
    visible_quarter_txts = [pd.Timestamp(q).strftime("%Y-%m-%d") for q in visible_quarters]
    ctx.derived.qa_export_expectation = {
        "metrics": ["QA_Buybacks"],
        "quarters": ([latest_visible_q_txt] if latest_visible_q_txt else []),
        "rows": _export_rows_snapshot(
            qa_checks,
            metrics=["QA_Buybacks"],
            quarters=([latest_visible_q] if pd.notna(latest_visible_q) else None),
        )
    }
    ctx.derived.needs_review_export_expectation = {
        "metrics": ["buybacks_cash"],
        "quarters": visible_quarter_txts,
        "rows": _export_rows_snapshot(
            needs_review,
            metrics=["buybacks_cash"],
            quarters=visible_quarters or None,
        )
    }


def finalize_workbook(ctx: WriterContext) -> None:
    state = ctx.state
    wb = ctx.wb

    def _set_defined_name(name: str, attr_text: str) -> None:
        txt = str(attr_text or "").strip()
        if not name or not txt:
            return
        try:
            if name in wb.defined_names:
                del wb.defined_names[name]
        except Exception:
            pass
        try:
            wb.defined_names.add(DefinedName(name=name, attr_text=txt))
        except Exception:
            try:
                wb.defined_names.append(DefinedName(name=name, attr_text=txt))
            except Exception:
                pass

    try:
        signals_base_df = ctx.derived.signals_base_df if ctx.derived.signals_base_df is not None else state["signals_base_df"]
        if "Hidden_Value_Audit" in wb.sheetnames and signals_base_df is not None and not signals_base_df.empty:
            ws_hva = wb["Hidden_Value_Audit"]
            if ws_hva.max_row >= 2:
                hdr = {str(c.value): i for i, c in enumerate(ws_hva[1], start=1) if c.value is not None}
                c_q = hdr.get("quarter")
                c_f = hdr.get("fcf_yield")
                c_out = hdr.get("output_value")
                c_pass = hdr.get("pass_fail")
                c_inputs = hdr.get("inputs_json")
                c_flag = hdr.get("flag_id")
                c_msg = hdr.get("qa_message")
                c_sev = hdr.get("qa_severity")
                if c_q and c_f:
                    sbase = signals_base_df.copy()
                    sbase["quarter"] = pd.to_datetime(sbase["quarter"], errors="coerce")
                    sbase = sbase[sbase["quarter"].notna()].copy()

                    def _num(v: Any) -> float | None:
                        try:
                            fv = float(v)
                            return fv if pd.notna(fv) else None
                        except Exception:
                            return None

                    by_q: Dict[pd.Timestamp, Dict[str, float | None]] = {}
                    for _, rr in sbase.iterrows():
                        qd = pd.Timestamp(rr["quarter"]).date()
                        by_q[qd] = {
                            "fcf_yield": _num(rr.get("fcf_yield")),
                            "fcf_ttm": _num(rr.get("fcf_ttm")),
                            "shares_out": _num(rr.get("shares_out")),
                        }

                    def _price_linked_fcf_yield_formula() -> str:
                        return "=IF(OR(Price=\"\",Price<=0,FCF_TTM=\"\",Shares=\"\",N(Shares)<=0),\"\",N(FCF_TTM)/(N(Price)*N(Shares)))"

                    for rix in range(2, ws_hva.max_row + 1):
                        qv = ws_hva.cell(row=rix, column=c_q).value
                        qd = pd.to_datetime(qv, errors="coerce")
                        if pd.isna(qd):
                            continue
                        rec = by_q.get(pd.Timestamp(qd).date())
                        if not rec:
                            continue

                        cell_f = ws_hva.cell(row=rix, column=c_f)
                        cur_f = _num(cell_f.value)
                        if cur_f is None:
                            fy = rec.get("fcf_yield")
                            if fy is not None:
                                cell_f.value = fy
                                cell_f.number_format = "0.0%"
                            else:
                                cell_f.value = _price_linked_fcf_yield_formula()
                                cell_f.number_format = "0.0%"
                                if c_msg:
                                    msg_cell = ws_hva.cell(row=rix, column=c_msg)
                                    msg = str(msg_cell.value or "")
                                    if "missing required inputs: fcf_yield" in msg.lower():
                                        msg_norm = re.sub(
                                            r"(?i)\s*missing required inputs:\s*fcf_yield\s*(\|\s*)?",
                                            "",
                                            msg,
                                        )
                                        msg_norm = re.sub(
                                            r"(?i)\s*fcf_yield needs market_cap or --price input \(used with shares\)\s*(\|\s*)?",
                                            "",
                                            msg_norm,
                                        )
                                        msg_norm = msg_norm.strip(" |")
                                        tail_note = "fcf_yield linked to Valuation Price input."
                                        msg_cell.value = (msg_norm + " | " + tail_note).strip(" |") if msg_norm else tail_note
                                if c_sev:
                                    sev_cell = ws_hva.cell(row=rix, column=c_sev)
                                    sev = str(sev_cell.value or "").strip().upper()
                                    if sev == "FAIL":
                                        sev_cell.value = "WARN"

                        if c_out and c_flag and c_inputs:
                            fid = str(ws_hva.cell(row=rix, column=c_flag).value or "").strip().upper()
                            if fid in {"C", "E"}:
                                out_cell = ws_hva.cell(row=rix, column=c_out)
                                if fid == "C":
                                    out_cell.value = "=IF(OR(FCF_TTM_Pos_Years=\"\",Pos_FCF_Ratio=\"\",FCF_Yield=\"\"),\"\",--AND(N(FCF_TTM_Pos_Years)>=1,N(Pos_FCF_Ratio)>=0.75,N(FCF_Yield)>=0.15))"
                                elif fid == "E":
                                    out_cell.value = "=IF(OR(Interest_Coverage=\"\",FCF_Yield=\"\"),\"\",--AND(N(Interest_Coverage)>=3,N(FCF_Yield)>=0.20))"
                                if c_pass:
                                    ws_hva.cell(row=rix, column=c_pass).value = f"=IF({get_column_letter(c_out)}{rix}=\"\",\"\",N({get_column_letter(c_out)}{rix})>=1)"

        if "Hidden_Value_Audit" in wb.sheetnames and "Hidden_Value_Flags" in wb.sheetnames:
            ws_hva = wb["Hidden_Value_Audit"]
            ws_hvf = wb["Hidden_Value_Flags"]
            if ws_hva.max_row >= 2 and ws_hvf.max_row >= 2:
                hdr_a = {str(c.value): i for i, c in enumerate(ws_hva[1], start=1) if c.value is not None}
                hdr_f = {str(c.value): i for i, c in enumerate(ws_hvf[1], start=1) if c.value is not None}
                c_flag = hdr_a.get("flag_id")
                c_q = hdr_a.get("quarter")
                c_out = hdr_a.get("output_value")
                c_msg = hdr_a.get("qa_message")
                c_inputs = hdr_a.get("inputs_json")
                c_name = hdr_a.get("flag_name")
                c_thr = hdr_a.get("threshold")
                c_f = hdr_a.get("fcf_yield")
                f_code = hdr_f.get("flag_code") or hdr_f.get("Flag")
                f_score = hdr_f.get("score")
                f_e1 = hdr_f.get("evidence_1")
                f_e2 = hdr_f.get("evidence_2")
                f_e3 = hdr_f.get("evidence_3")
                f_visible_support = hdr_f.get("visible_support")
                audit_row_by_flag = {}
                if c_flag:
                    for rr in range(2, ws_hva.max_row + 1):
                        fid = str(ws_hva.cell(row=rr, column=c_flag).value or "").strip().upper()
                        if fid:
                            audit_row_by_flag[fid] = rr
                if f_code and f_score and f_e1 and f_e2 and f_e3:
                    for rr in range(2, ws_hvf.max_row + 1):
                        fid = str(ws_hvf.cell(row=rr, column=f_code).value or "").strip().upper()
                        if fid not in audit_row_by_flag:
                            continue
                        ar = audit_row_by_flag[fid]
                        out_ref = f"'Hidden_Value_Audit'!${get_column_letter(c_out)}${ar}"
                        code_cell = f"${get_column_letter(f_code)}{rr}"
                        q_ref = f"'Hidden_Value_Audit'!${get_column_letter(c_q)}${ar}"
                        msg_ref = f"'Hidden_Value_Audit'!${get_column_letter(c_msg)}${ar}" if c_msg else "\"\""
                        in_ref = f"'Hidden_Value_Audit'!${get_column_letter(c_inputs)}${ar}" if c_inputs else "\"\""
                        name_ref = f"'Hidden_Value_Audit'!${get_column_letter(c_name)}${ar}" if c_name else code_cell
                        thr_ref = f"'Hidden_Value_Audit'!${get_column_letter(c_thr)}${ar}" if c_thr else "\"\""
                        ws_hvf.cell(row=rr, column=f_score).value = f"=IF({out_ref}=\"\",\"\",IF(N({out_ref})>=1,100,0))"
                        e1_formula = f"=IF({code_cell}=\"\",\"\",{name_ref}&\" | Quarter: \"&TEXT({q_ref},\"yyyy-mm-dd\"))"
                        e2_formula = f"=IF({code_cell}=\"\",\"\",\"Threshold: \"&{thr_ref}&IF({msg_ref}<>\"\",\" | \"&{msg_ref},\"\"))"
                        e3_formula = f"=IF({code_cell}=\"\",\"\",\"Inputs: \"&{in_ref})"
                        if fid == "C" and c_f:
                            fy_ref = f"'Hidden_Value_Audit'!${get_column_letter(c_f)}${ar}"
                            e2_formula = (
                                f"=IF({code_cell}=\"\",\"\",IF({fy_ref}=\"\",\"Price-linked via Valuation input | Status: WAIT\","
                                f"IF(N({out_ref})>=1,\"Price-linked via Valuation input | Status: PASS\",\"Price-linked via Valuation input | Status: FAIL\")))"
                            )
                            e3_formula = f"=IF({code_cell}=\"\",\"\",\"Trigger active from Hidden_Value_Audit C\")"
                        elif fid == "E" and c_f:
                            fy_ref = f"'Hidden_Value_Audit'!${get_column_letter(c_f)}${ar}"
                            e2_formula = (
                                f"=IF({code_cell}=\"\",\"\",IF({fy_ref}=\"\",\"Price-linked via Valuation input | Status: WAIT\","
                                f"IF(N({out_ref})>=1,\"Price-linked via Valuation input | Status: PASS\",\"Price-linked via Valuation input | Status: FAIL\")))"
                            )
                            e3_formula = f"=IF({code_cell}=\"\",\"\",\"Trigger active from Hidden_Value_Audit E\")"
                        ws_hvf.cell(row=rr, column=f_e1).value = e1_formula
                        ws_hvf.cell(row=rr, column=f_e2).value = e2_formula
                        ws_hvf.cell(row=rr, column=f_e3).value = e3_formula
                        if f_visible_support:
                            if fid == "C":
                                ws_hvf.cell(row=rr, column=f_visible_support).value = (
                                    "=IF(OR(FCF_TTM_Pos_Years=\"\",Pos_FCF_Ratio=\"\",FCF_Yield=\"\"),"
                                    "\"(price-linked)\","
                                    "\"Positive FCF years \"&TEXT(FCF_TTM_Pos_Years,\"0\")&"
                                    "\", Positive FCF ratio \"&TEXT(Pos_FCF_Ratio,\"0%\")&"
                                    "\", FCF yield \"&TEXT(FCF_Yield,\"0.0%\")&"
                                    "\" (price-linked)\")"
                                )
                            elif fid == "E":
                                ws_hvf.cell(row=rr, column=f_visible_support).value = (
                                    "=IF(OR(Interest_Coverage=\"\",FCF_Yield=\"\"),"
                                    "\"(price-linked)\","
                                    "\"Interest cover \"&TEXT(Interest_Coverage,\"0.00x\")&"
                                    "\", FCF yield \"&TEXT(FCF_Yield,\"0.0%\")&"
                                    "\" (price-linked)\")"
                                )

                    row_end_col = get_column_letter(ws_hvf.max_column)
                    ws_hvf.conditional_formatting.add(
                        f"A2:{row_end_col}{ws_hvf.max_row}",
                        FormulaRule(
                            formula=["OR($B2=\"C\",$B2=\"E\")"],
                            fill=PatternFill("solid", fgColor="E2F0D9"),
                        ),
                    )
                    ws_hvf.conditional_formatting.add(
                        f"D2:D{ws_hvf.max_row}",
                        CellIsRule(operator="greaterThanOrEqual", formula=["70"], fill=PatternFill("solid", fgColor="C6EFCE")),
                    )
                    ws_hvf.conditional_formatting.add(
                        f"D2:D{ws_hvf.max_row}",
                        CellIsRule(operator="between", formula=["40", "69.999"], fill=PatternFill("solid", fgColor="FFEB9C")),
                    )
                    ws_hvf.conditional_formatting.add(
                        f"D2:D{ws_hvf.max_row}",
                        CellIsRule(operator="lessThan", formula=["40"], fill=PatternFill("solid", fgColor="FFC7CE")),
                    )
                    for rr in range(2, ws_hvf.max_row + 1):
                        if str(ws_hvf.cell(row=rr, column=f_code).value or "").strip() != "":
                            ws_hvf.row_dimensions[rr].height = 32

        if "Valuation" in wb.sheetnames and "Hidden_Value_Audit" in wb.sheetnames and "Hidden_Value_Flags" in wb.sheetnames:
            ws_val = wb["Valuation"]
            if "Promise_Progress_UI" in wb.sheetnames:
                ws_pp = wb["Promise_Progress_UI"]

                def _best_progress_latest_local(metric_name: str) -> str:
                    best_txt = ""
                    best_score: Tuple[int, int] = (-1, -1)
                    current_block_ord = -1
                    for rr in range(1, ws_pp.max_row + 1):
                        header_txt = str(ws_pp.cell(row=rr, column=1).value or "").strip()
                        m_header = re.match(r"Promise progress \(As of (\d{4}-\d{2}-\d{2})\)$", header_txt)
                        if m_header:
                            try:
                                q_ts = pd.Timestamp(m_header.group(1))
                                current_block_ord = int(q_ts.year) * 4 + (((int(q_ts.month) - 1) // 3) + 1)
                            except Exception:
                                current_block_ord = -1
                            continue
                        metric_txt = str(ws_pp.cell(row=rr, column=1).value or "").strip()
                        if metric_txt != metric_name:
                            continue
                        latest_txt = str(ws_pp.cell(row=rr, column=3).value or "").strip()
                        if latest_txt.lower() in {"", "not yet measurable", "nan", "none", "null"}:
                            continue
                        score_tuple = (current_block_ord, len(latest_txt))
                        if score_tuple > best_score:
                            best_score = score_tuple
                            best_txt = latest_txt
                    return best_txt

                valuation_row = 1
                while valuation_row <= min(ws_val.max_row, 120):
                    metric_txt = str(ws_val.cell(row=valuation_row, column=15).value or "").strip()
                    stated_txt = str(ws_val.cell(row=valuation_row, column=17).value or "").strip().lower()
                    applies_txt = str(ws_val.cell(row=valuation_row, column=18).value or "").strip().lower()
                    guidance_txt = str(ws_val.cell(row=valuation_row, column=19).value or "").strip().lower()
                    if metric_txt == "Cost savings" and (
                        "raised target" in guidance_txt
                        or "annualized savings" in guidance_txt
                        or applies_txt in {"run-rate", "annualized"}
                        or "carry-fwd" in stated_txt
                    ):
                        best_latest = _best_progress_latest_local("Cost savings target")
                        if best_latest:
                            best_latest = re.sub(r"\s+run-rate\b", "", best_latest, flags=re.I).strip()
                            ws_val.cell(row=valuation_row, column=26, value=f"{best_latest} realized").alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    elif metric_txt == "PB Bank liquidity" and (
                        "bank-held leases" in guidance_txt
                        or "liquidity release" in guidance_txt
                        or "carry-fwd" in stated_txt
                    ):
                        best_latest = _best_progress_latest_local("PB Bank liquidity release")
                        if best_latest:
                            ws_val.cell(row=valuation_row, column=26, value=f"{best_latest} realized").alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    valuation_row += 1
        if "Hidden_Value_Base" in wb.sheetnames:
            ws_hvb = wb["Hidden_Value_Base"]
            if ws_hvb.max_row >= 2:
                hdr_b = {
                    str(c.value): i
                    for i, c in enumerate(ws_hvb[1], start=1)
                    if c.value is not None and str(c.value).strip() != ""
                }
                latest_row = 0
                for rr in range(ws_hvb.max_row, 1, -1):
                    if any(ws_hvb.cell(row=rr, column=cc).value not in (None, "") for cc in hdr_b.values()):
                        latest_row = rr
                        break
                if latest_row >= 2:
                    def _base_ref(header_name: str) -> str:
                        col_idx = hdr_b.get(header_name)
                        if not col_idx:
                            return ""
                        return f"'Hidden_Value_Base'!${get_column_letter(col_idx)}${latest_row}"

                    hv_fcf_formula = "Equity_FCF_Yield" if "Equity_FCF_Yield" in wb.defined_names else _base_ref("fcf_yield")
                    if hv_fcf_formula:
                        _set_defined_name("FCF_Yield", hv_fcf_formula)
                    hv_ttm_pos_ref = _base_ref("fcf_ttm_pos_years")
                    if hv_ttm_pos_ref:
                        _set_defined_name("FCF_TTM_Pos_Years", hv_ttm_pos_ref)
                    hv_pos_ratio_ref = _base_ref("pos_fcf_ratio")
                    if hv_pos_ratio_ref:
                        _set_defined_name("Pos_FCF_Ratio", hv_pos_ratio_ref)
                    hv_interest_ref = _base_ref("interest_coverage") or _base_ref("interest_coverage_cash") or _base_ref("interest_coverage_pnl")
                    if hv_interest_ref:
                        _set_defined_name("Interest_Coverage", hv_interest_ref)
    except Exception:
        try:
            import traceback

            print("[hidden_flags_sync] WARN row-link failed", flush=True)
            traceback.print_exc()
        except Exception:
            pass

    try:
        current = list(wb.sheetnames)
        ordered = [s for s in ctx.desired_sheet_order if s in current] + [s for s in current if s not in ctx.desired_sheet_order]
        for target_idx, name in enumerate(ordered):
            ws_obj = wb[name]
            cur_idx = wb._sheets.index(ws_obj)
            if cur_idx != target_idx:
                wb._sheets.insert(target_idx, wb._sheets.pop(cur_idx))
    except Exception:
        pass

    try:
        if "History_Q" in wb.sheetnames:
            insert_at = wb.sheetnames.index("History_Q")
            for name in ctx.raw_sheet_cluster:
                if name not in wb.sheetnames:
                    continue
                ws_obj = wb[name]
                cur_idx = wb._sheets.index(ws_obj)
                if cur_idx != insert_at:
                    wb._sheets.insert(insert_at, wb._sheets.pop(cur_idx))
                insert_at += 1
    except Exception:
        pass

    try:
        ws_sum = wb["SUMMARY"]
        for col_name, width in {
            "A": 42.0,
            "B": 18.0,
            "C": 14.0,
            "D": 12.0,
            "E": 12.0,
            "F": 12.0,
            "G": 12.0,
            "H": 12.0,
        }.items():
            ws_sum.column_dimensions[col_name].width = max(width, min(width, ws_sum.column_dimensions[col_name].width or width))
    except Exception:
        pass

    date_cols = {"quarter", "start", "end", "filed", "period_end", "period_start", "filed_date"}
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        headers = [c.value for c in ws[1]]
        for idx, header in enumerate(headers, start=1):
            if header in date_cols:
                for cell in ws[get_column_letter(idx)][1:]:
                    if cell.value is None:
                        continue
                    cell.number_format = "yyyy-mm-dd"
                ws.column_dimensions[get_column_letter(idx)].width = max(
                    12,
                    ws.column_dimensions[get_column_letter(idx)].width or 12,
                )

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    wb.save(ctx.data.out_path)
