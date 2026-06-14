from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from typing import Any, Dict, List, Mapping, MutableMapping, Optional, Tuple

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .derivative_oci_bridge import DERIVATIVE_EXPOSURE_COLUMNS


@dataclass(frozen=True)
class DerivativeOciBridgeRenderDeps:
    runtime: MutableMapping[str, Any]


def write_derivative_oci_bridge_sheet(
    deps: DerivativeOciBridgeRenderDeps,
    bridge_df: Any,
    exposure_df: Any = None,
) -> None:
    """Write the derivative accounting source sheet and memo diagnostics.

    The first table is the normalized extraction output. Everything appended
    below it is workbook UI for auditability and diagnostics only; none of
    these rows feed reported actuals, valuation, or the production GPRE
    crush proxy.
    """
    runtime = deps.runtime
    wb = runtime["wb"]
    _write_sheet = runtime["_write_sheet"]
    _safe_cell = runtime["_safe_cell"]
    _get_analysis_sheet_style_bundle = runtime["_get_analysis_sheet_style_bundle"]
    font_size = runtime["font_size"]
    header_size = runtime["header_size"]
    ctx_ref = runtime.get("ctx_ref")
    operating_driver_history_rows = runtime.get("operating_driver_history_rows")
    _write_sheet("Derivative_OCI_Bridge", bridge_df)
    ws = wb["Derivative_OCI_Bridge"]
    exposure = exposure_df.copy() if isinstance(exposure_df, pd.DataFrame) else pd.DataFrame()
    for col in DERIVATIVE_EXPOSURE_COLUMNS:
        if col not in exposure.columns:
            exposure[col] = pd.NA
    exposure = exposure[DERIVATIVE_EXPOSURE_COLUMNS]

    theme = _get_analysis_sheet_style_bundle()
    title_fill_local = copy(theme["title_fill"])
    section_fill_local = copy(theme["section_fill"])
    header_fill_local = copy(theme["header_fill"])
    neutral_fill = copy(theme["neutral_fill_alt"])
    alt_fill = copy(theme["neutral_fill"])
    thin_border_local = copy(theme["thin_border"])
    title_font_local = Font(bold=True, size=header_size + 2, color="FFFFFF")
    subtitle_font = Font(size=font_size, italic=True, color=str(theme["text_muted"]))
    header_font_local = Font(bold=True, size=header_size, color=str(theme["text_dark"]))
    body_font_local = Font(size=font_size, color=str(theme["text_dark"]))

    start_row = ws.max_row + 6
    headers = list(DERIVATIVE_EXPOSURE_COLUMNS)
    last_col = len(headers)
    last_col_letter = get_column_letter(last_col)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=last_col)
    title_cell = ws.cell(row=start_row, column=1, value="Open Derivative Position Exposure")
    title_cell.fill = title_fill_local
    title_cell.font = title_font_local
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    for cc in range(1, last_col + 1):
        cell = ws.cell(row=start_row, column=cc)
        cell.fill = copy(title_fill_local)
        cell.border = copy(thin_border_local)
    ws.row_dimensions[start_row].height = 24.0

    subtitle_row = start_row + 1
    ws.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=last_col)
    subtitle_cell = ws.cell(
        row=subtitle_row,
        column=1,
        value=(
            "Notional exposure by commodity, instrument, accounting treatment and likely P&L line. "
            "For GPRE, notional amounts are disclosed in thousands of units. Notional exposure does not equal "
            "fair value, OCI, AOCI or derivative P&L."
        ),
    )
    subtitle_cell.fill = section_fill_local
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for cc in range(1, last_col + 1):
        cell = ws.cell(row=subtitle_row, column=cc)
        cell.fill = copy(section_fill_local)
        cell.border = copy(thin_border_local)
    ws.row_dimensions[subtitle_row].height = 30.0

    header_row = start_row + 3
    for cc, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=cc, value=header)
        cell.fill = copy(header_fill_local)
        cell.font = copy(header_font_local)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(thin_border_local)
    ws.row_dimensions[header_row].height = 28.0

    data_start = header_row + 1
    if exposure.empty:
        ws.cell(row=data_start, column=1, value="No disclosed open derivative position notional data found.")
        for cc in range(1, last_col + 1):
            cell = ws.cell(row=data_start, column=cc)
            cell.fill = copy(neutral_fill)
            cell.border = copy(thin_border_local)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        data_end = data_start
    else:
        for rr, (_, rec) in enumerate(exposure.iterrows(), start=data_start):
            fill = neutral_fill if (rr - data_start) % 2 == 0 else alt_fill
            for cc, header in enumerate(headers, start=1):
                value = rec.get(header)
                value = None if pd.isna(value) else _safe_cell(value)
                cell = ws.cell(row=rr, column=cc, value=value)
                cell.fill = copy(fill)
                cell.font = copy(body_font_local)
                cell.border = copy(thin_border_local)
                cell.alignment = Alignment(
                    horizontal="right" if header in {"Long notional", "Short notional", "Net notional"} else "left",
                    vertical="center",
                    wrap_text=header in {"Interpretation", "Source / note", "Scale"},
                )
                if header == "Quarter":
                    cell.number_format = "yyyy-mm-dd"
                elif header in {"Long notional", "Short notional", "Net notional"}:
                    cell.number_format = '#,##0.0;(#,##0.0);-'
            ws.row_dimensions[rr].height = 22.0
        data_end = data_start + len(exposure) - 1

    width_map = {
        "A": 14,
        "B": 18,
        "C": 28,
        "D": 24,
        "E": 13,
        "F": 16,
        "G": 16,
        "H": 16,
        "I": 12,
        "J": 14,
        "K": 18,
        "L": 54,
        "M": 48,
    }
    for letter, width in width_map.items():
        ws.column_dimensions[letter].width = max(float(ws.column_dimensions[letter].width or 0), float(width))

    if data_end >= data_start and not exposure.empty:
        ref = f"A{header_row}:{last_col_letter}{data_end}"
        try:
            t = Table(displayName="DerivativePositionExposure", ref=ref)
            t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(t)
        except Exception:
            pass
        net_col_letter = get_column_letter(headers.index("Net notional") + 1)
        net_range = f"{net_col_letter}{data_start}:{net_col_letter}{data_end}"
        ws.conditional_formatting.add(
            net_range,
            CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="E2F0D9")),
        )
        ws.conditional_formatting.add(
            net_range,
            CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FCE4D6")),
        )

    note_row = data_end + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=last_col)
    note_cell = ws.cell(
        row=note_row,
        column=1,
        value=(
            "Note: Notional exposure is shown on the same scale as disclosed by the company. For GPRE, notional "
            "amounts are disclosed in thousands of units. Notional exposure does not equal fair value, OCI, "
            "AOCI or derivative P&L. Fair value depends on entry price, market price, timing, contract terms "
            "and hedge designation. Net derivative asset/liability is a period-end balance-sheet snapshot, "
            "while derivative P&L shows income-statement impact during the period."
        ),
    )
    note_cell.fill = copy(section_fill_local)
    note_cell.font = copy(body_font_local)
    note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for cc in range(1, last_col + 1):
        cell = ws.cell(row=note_row, column=cc)
        cell.fill = copy(section_fill_local)
        cell.border = copy(thin_border_local)
    ws.row_dimensions[note_row].height = 42.0

    def _diagnostic_quarter(value: Any) -> Optional[date]:
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            return None
        return ts.date()

    def _diagnostic_usd_to_millions(value: Any) -> Optional[float]:
        num = pd.to_numeric(value, errors="coerce")
        if pd.isna(num):
            return None
        return float(num) / 1_000_000.0

    def _diagnostic_float(value: Any) -> Optional[float]:
        num = pd.to_numeric(value, errors="coerce")
        if pd.isna(num):
            return None
        return float(num)

    gallons_denominator_by_quarter: Dict[date, Tuple[Optional[float], str]] = {}
    driver_history_rows_for_diagnostics: Any = operating_driver_history_rows
    if not driver_history_rows_for_diagnostics and ctx_ref is not None:
        try:
            driver_history_rows_for_diagnostics = ctx_ref.data.operating_driver_history_rows
        except Exception:
            driver_history_rows_for_diagnostics = []
    if isinstance(driver_history_rows_for_diagnostics, list):
        for wanted_key, label_txt in (
            ("ethanol_gallons_produced", "Ethanol gallons produced"),
            ("ethanol_gallons_sold", "Ethanol gallons sold"),
        ):
            for rec in driver_history_rows_for_diagnostics:
                if str(rec.get("_driver_key") or "").strip() != wanted_key:
                    continue
                qd = _diagnostic_quarter(rec.get("Quarter"))
                if qd is None or qd in gallons_denominator_by_quarter:
                    continue
                val_m = _diagnostic_float(rec.get("Value"))
                if val_m is None or abs(val_m) < 1e-9:
                    continue
                gallons_denominator_by_quarter[qd] = (float(val_m), label_txt)

    def _diagnostic_denominator(qd: Optional[date]) -> Tuple[Optional[str], Optional[float], Optional[float], str]:
        if qd is None:
            return None, None, None, "denominator not available"
        gallons_m, label_txt = gallons_denominator_by_quarter.get(qd, (None, "denominator not available"))
        if gallons_m is None or abs(float(gallons_m)) < 1e-9:
            return "denominator not available", None, None, "denominator not available"
        return label_txt, float(gallons_m), None, ""

    def _amount_per_gallon(amount_m: Optional[float], gallons_m: Optional[float]) -> Optional[float]:
        if amount_m is None or gallons_m is None or abs(float(gallons_m)) < 1e-9:
            return None
        return float(amount_m) / float(gallons_m)

    pnl_specs = (
        ("Total derivative P&L", "derivative_gain_loss_pnl_total_usd", "Revenue/COGS", "Current-quarter reported margin impact"),
        ("Derivative P&L in revenue", "derivative_gain_loss_revenue_usd", "Revenue", "Revenue-side hedge/derivative impact"),
        ("Derivative P&L in COGS", "derivative_gain_loss_cogs_usd", "COGS", "COGS-side hedge/derivative impact"),
        ("Cash-flow hedge reclass to P&L", "cash_flow_hedge_reclass_total_usd", "Revenue/COGS", "AOCI reclass now included in P&L"),
        ("Fair-value hedge P&L", "fair_value_hedge_total_pnl_usd", "COGS", "Fair-value hedge impact recognized in P&L"),
        ("Non-designated derivative P&L", "non_designated_derivative_pnl_total_usd", "Revenue/COGS", "Economic/non-designated MTM recognized in P&L"),
    )
    pnl_per_gallon_rows: List[Dict[str, Any]] = []
    def _append_pnl_per_gallon_row(
        rows_out: List[Dict[str, Any]],
        *,
        qd: Optional[date],
        metric_txt: str,
        amount_m: Optional[float],
        pnl_line: str,
        interpretation: str,
        denom_label: Optional[str],
        gallons_m: Optional[float],
        denom_warning: str,
    ) -> None:
        if amount_m is None:
            return
        final_interpretation = interpretation if not denom_warning else f"{interpretation}; {denom_warning}"
        rows_out.append(
            {
                "Quarter": pd.Timestamp(qd) if qd is not None else pd.NaT,
                "Metric": metric_txt,
                "Amount ($m)": amount_m,
                "Denominator": denom_label or "denominator not available",
                "Gallons (m)": gallons_m,
                "$/gal": _amount_per_gallon(amount_m, gallons_m),
                "P&L line": pnl_line,
                "Interpretation": final_interpretation,
            }
        )

    if isinstance(bridge_df, pd.DataFrame) and not bridge_df.empty:
        for _, rec in bridge_df.iterrows():
            qd = _diagnostic_quarter(rec.get("quarter"))
            denom_label, gallons_m, _, denom_warning = _diagnostic_denominator(qd)
            for metric_txt, field_name, pnl_line, interpretation in pnl_specs:
                amount_m = _diagnostic_usd_to_millions(rec.get(field_name))
                _append_pnl_per_gallon_row(
                    pnl_per_gallon_rows,
                    qd=qd,
                    metric_txt=metric_txt,
                    amount_m=amount_m,
                    pnl_line=pnl_line,
                    interpretation=interpretation,
                    denom_label=denom_label,
                    gallons_m=gallons_m,
                    denom_warning=denom_warning,
                )
            total_m = _diagnostic_usd_to_millions(rec.get("derivative_gain_loss_pnl_total_usd"))
            component_values = [
                _diagnostic_usd_to_millions(rec.get("non_designated_derivative_pnl_total_usd")),
                _diagnostic_usd_to_millions(rec.get("cash_flow_hedge_reclass_total_usd")),
                _diagnostic_usd_to_millions(rec.get("fair_value_hedge_total_pnl_usd")),
            ]
            available_components = [float(x) for x in component_values if x is not None]
            if total_m is not None and available_components:
                residual_m = float(total_m) - sum(available_components)
                if abs(float(residual_m)) < 0.0005:
                    residual_m = 0.0
                _append_pnl_per_gallon_row(
                    pnl_per_gallon_rows,
                    qd=qd,
                    metric_txt="P&L component residual / unallocated",
                    amount_m=residual_m,
                    pnl_line="Unallocated",
                    interpretation="Residual from incomplete component disclosure, quarterization, rounding, or unallocated derivative/hedge components.",
                    denom_label=denom_label,
                    gallons_m=gallons_m,
                    denom_warning=denom_warning,
                )

    deferred_specs = (
        ("Net derivative asset/liability", "derivative_net_asset_liability_usd", "Balance sheet snapshot", "Not current-quarter P&L; open derivative fair-value snapshot"),
        ("Derivative OCI movement", "derivative_oci_current_period_usd", "Current-period OCI, not P&L", "Not current-quarter P&L; unrealized cash-flow hedge movement in equity/AOCI"),
        ("Derivative AOCI", "derivative_aoci_ending_balance_usd", "Accumulated equity balance", "Not current-quarter P&L; deferred cash-flow hedge amount in equity"),
        ("AOCI reclass to earnings, net of tax/sign convention", "derivative_aoci_reclassified_to_earnings_usd", "OCI/AOCI statement presentation", "Not used as current-quarter margin impact here; separate from pre-tax P&L reclass"),
    )
    deferred_rows: List[Dict[str, Any]] = []
    if isinstance(bridge_df, pd.DataFrame) and not bridge_df.empty:
        for _, rec in bridge_df.iterrows():
            qd = _diagnostic_quarter(rec.get("quarter"))
            denom_label, gallons_m, _, denom_warning = _diagnostic_denominator(qd)
            for metric_txt, field_name, accounting_status, interpretation in deferred_specs:
                amount_m = _diagnostic_usd_to_millions(rec.get(field_name))
                if amount_m is None:
                    continue
                final_interpretation = interpretation if not denom_warning else f"{interpretation}; diagnostic scaling only; {denom_warning}"
                deferred_rows.append(
                    {
                        "Quarter": pd.Timestamp(qd) if qd is not None else pd.NaT,
                        "Metric": metric_txt,
                        "Amount ($m)": amount_m,
                        "Denominator": denom_label or "denominator not available",
                        "Gallons (m)": gallons_m,
                        "$/gal": _amount_per_gallon(amount_m, gallons_m),
                        "Accounting status": accounting_status,
                        "Interpretation": final_interpretation,
                    }
                )

    def _margin_bucket_for_commodity(commodity: Any) -> str:
        low = str(commodity or "").strip().lower()
        if "corn oil" in low or "renewable corn oil" in low:
            return "Coproduct output"
        if "distiller" in low:
            return "Coproduct output"
        if "natural gas" in low:
            return "Production energy input"
        if low == "corn" or "corn" in low:
            return "Core crush input"
        if "ethanol" in low:
            return "Core crush output"
        return "Other / not classified"

    def _margin_bucket_interpretation(bucket: str) -> str:
        if bucket == "Core crush input":
            return "Input hedge: affects corn cost / COGS economics; exposure classification only; fair value by commodity not disclosed"
        if bucket == "Production energy input":
            return "Input hedge: affects production energy / COGS economics; exposure classification only; fair value by commodity not disclosed"
        if bucket == "Core crush output":
            return "Output hedge: affects ethanol revenue economics; exposure classification only; fair value by commodity not disclosed"
        if bucket == "Coproduct output":
            return "Coproduct hedge: affects coproduct revenue economics; exposure classification only; fair value by commodity not disclosed"
        return "Exposure classification only; fair value by commodity not disclosed"

    margin_bucket_rows: List[Dict[str, Any]] = []
    if isinstance(exposure, pd.DataFrame) and not exposure.empty:
        for _, rec in exposure.iterrows():
            bucket = _margin_bucket_for_commodity(rec.get("Commodity"))
            margin_bucket_rows.append(
                {
                    "Quarter": rec.get("Quarter"),
                    "Commodity": rec.get("Commodity"),
                    "Instrument": rec.get("Instrument"),
                    "Accounting bucket": rec.get("Accounting bucket"),
                    "Direction": rec.get("Direction"),
                    "Net notional": rec.get("Net notional"),
                    "Unit": rec.get("Unit"),
                    "Scale": rec.get("Scale"),
                    "Likely P&L line": rec.get("Likely P&L line"),
                    "Margin bucket": bucket,
                    "Interpretation": _margin_bucket_interpretation(bucket),
                }
            )

    def _quarterly_label(qd_in: Any) -> Optional[str]:
        qd = _diagnostic_quarter(qd_in)
        if qd is None:
            return None
        try:
            period = pd.Timestamp(qd).to_period("Q")
            return f"{int(period.year)}-Q{int(period.quarter)}"
        except Exception:
            return str(qd)

    pnl_rows_by_metric_quarter: Dict[Tuple[str, str], Dict[str, Any]] = {}
    quarter_labels_in_order: List[str] = []
    quarter_sort_key: Dict[str, Any] = {}
    for rec in pnl_per_gallon_rows:
        q_label = _quarterly_label(rec.get("Quarter"))
        metric_txt = str(rec.get("Metric") or "").strip()
        if not q_label or not metric_txt:
            continue
        pnl_rows_by_metric_quarter[(metric_txt, q_label)] = rec
        if q_label not in quarter_sort_key:
            quarter_sort_key[q_label] = pd.to_datetime(rec.get("Quarter"), errors="coerce")
            quarter_labels_in_order.append(q_label)
    quarter_labels_in_order = sorted(
        quarter_labels_in_order,
        key=lambda label: quarter_sort_key.get(label) if not pd.isna(quarter_sort_key.get(label)) else pd.Timestamp.max,
    )
    quarterly_metric_specs = (
        ("Total derivative P&L / gal", "Total derivative P&L"),
        ("Derivative P&L in revenue / gal", "Derivative P&L in revenue"),
        ("Derivative P&L in COGS / gal", "Derivative P&L in COGS"),
        ("Cash-flow hedge reclass / gal", "Cash-flow hedge reclass to P&L"),
        ("Fair-value hedge P&L / gal", "Fair-value hedge P&L"),
        ("Non-designated derivative P&L / gal", "Non-designated derivative P&L"),
        ("P&L component residual / unallocated / gal", "P&L component residual / unallocated"),
    )
    quarterly_impact_rows: List[Dict[str, Any]] = []
    for display_metric, source_metric in quarterly_metric_specs:
        row: Dict[str, Any] = {"Metric": display_metric}
        for q_label in quarter_labels_in_order:
            row[q_label] = pnl_rows_by_metric_quarter.get((source_metric, q_label), {}).get("$/gal")
        quarterly_impact_rows.append(row)
    denominator_label_by_quarter: Dict[str, str] = {}
    denominator_value_by_quarter: Dict[str, Optional[float]] = {}
    for rec in pnl_per_gallon_rows:
        q_label = _quarterly_label(rec.get("Quarter"))
        if not q_label:
            continue
        if q_label not in denominator_value_by_quarter:
            denominator_value_by_quarter[q_label] = _diagnostic_float(rec.get("Gallons (m)"))
            denominator_label_by_quarter[q_label] = str(rec.get("Denominator") or "Ethanol gallons denominator").strip()
    denominator_metric = "Ethanol gallons produced (m)"
    if denominator_label_by_quarter and not any(str(v).lower() == "ethanol gallons produced" for v in denominator_label_by_quarter.values()):
        denominator_metric = "Ethanol gallons denominator (m)"
    denominator_row: Dict[str, Any] = {"Metric": denominator_metric}
    for q_label in quarter_labels_in_order:
        denominator_row[q_label] = denominator_value_by_quarter.get(q_label)
    quarterly_impact_rows.append(denominator_row)

    def _write_diagnostic_section_title(row_num: int, title: str, subtitle: str, end_col: int) -> int:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=end_col)
        cell = ws.cell(row=row_num, column=1, value=title)
        cell.fill = copy(title_fill_local)
        cell.font = copy(title_font_local)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for cc in range(1, end_col + 1):
            ws.cell(row=row_num, column=cc).fill = copy(title_fill_local)
            ws.cell(row=row_num, column=cc).border = copy(thin_border_local)
        ws.row_dimensions[row_num].height = 24.0
        row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=end_col)
        sub_cell = ws.cell(row=row_num, column=1, value=subtitle)
        sub_cell.fill = copy(section_fill_local)
        sub_cell.font = subtitle_font
        sub_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for cc in range(1, end_col + 1):
            ws.cell(row=row_num, column=cc).fill = copy(section_fill_local)
            ws.cell(row=row_num, column=cc).border = copy(thin_border_local)
        ws.row_dimensions[row_num].height = 36.0
        return row_num + 2

    def _diagnostic_table_name(title: str) -> str:
        return re.sub(r"[^A-Za-z0-9]+", "", title)[:240] or "DerivativeDiagnosticTable"

    def _write_diagnostic_table(row_num: int, title: str, headers_in: List[str], rows_in: List[Dict[str, Any]], table_name: str) -> int:
        end_col = len(headers_in)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=end_col)
        title_cell = ws.cell(row=row_num, column=1, value=title)
        title_cell.fill = copy(section_fill_local)
        title_cell.font = copy(header_font_local)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        for cc in range(1, end_col + 1):
            ws.cell(row=row_num, column=cc).fill = copy(section_fill_local)
            ws.cell(row=row_num, column=cc).border = copy(thin_border_local)
        ws.row_dimensions[row_num].height = 24.0
        header_row_num = row_num + 1
        for cc, header in enumerate(headers_in, start=1):
            cell = ws.cell(row=header_row_num, column=cc, value=header)
            cell.fill = copy(header_fill_local)
            cell.font = copy(header_font_local)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = copy(thin_border_local)
        ws.row_dimensions[header_row_num].height = 28.0
        data_start_row = header_row_num + 1
        if not rows_in:
            rows_in = [{headers_in[0]: "No disclosed data available."}]
        for rr, rec in enumerate(rows_in, start=data_start_row):
            fill = neutral_fill if (rr - data_start_row) % 2 == 0 else alt_fill
            for cc, header in enumerate(headers_in, start=1):
                value = rec.get(header)
                value = None if pd.isna(value) else _safe_cell(value)
                cell = ws.cell(row=rr, column=cc, value=value)
                cell.fill = copy(fill)
                cell.font = copy(body_font_local)
                cell.border = copy(thin_border_local)
                cell.alignment = Alignment(
                    horizontal="right" if header in {"Amount ($m)", "Gallons (m)", "$/gal", "Net notional"} else "left",
                    vertical="center",
                    wrap_text=header in {"Interpretation", "P&L line", "Accounting status", "Margin bucket"},
                )
                if header == "Quarter":
                    cell.number_format = "yyyy-mm-dd"
                elif header == "Amount ($m)":
                    cell.number_format = '#,##0.0;(#,##0.0);-'
                elif header == "Gallons (m)":
                    cell.number_format = '#,##0.0'
                elif header == "$/gal":
                    cell.number_format = '$0.000;($0.000);-'
                elif header == "Net notional":
                    cell.number_format = '#,##0.0;(#,##0.0);-'
            ws.row_dimensions[rr].height = 22.0
        data_end_row = data_start_row + len(rows_in) - 1
        try:
            table_ref = f"A{header_row_num}:{get_column_letter(end_col)}{data_end_row}"
            t = Table(displayName=table_name, ref=table_ref)
            t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(t)
        except Exception:
            pass
        return data_end_row + 2

    def _write_quarterly_impact_table(
        row_num: int,
        title: str,
        headers_in: List[str],
        rows_in: List[Dict[str, Any]],
    ) -> int:
        end_col = max(2, len(headers_in))
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=end_col)
        title_cell = ws.cell(row=row_num, column=1, value=title)
        title_cell.fill = copy(section_fill_local)
        title_cell.font = copy(header_font_local)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        for cc in range(1, end_col + 1):
            ws.cell(row=row_num, column=cc).fill = copy(section_fill_local)
            ws.cell(row=row_num, column=cc).border = copy(thin_border_local)
        ws.row_dimensions[row_num].height = 24.0
        header_row_num = row_num + 1
        for cc, header in enumerate(headers_in, start=1):
            cell = ws.cell(row=header_row_num, column=cc, value=header)
            cell.fill = copy(header_fill_local)
            cell.font = copy(header_font_local)
            cell.border = copy(thin_border_local)
            cell.alignment = Alignment(horizontal="center" if cc > 1 else "left", vertical="center", wrap_text=True)
        ws.row_dimensions[header_row_num].height = 28.0
        data_start_row = header_row_num + 1
        if not rows_in:
            rows_in = [{"Metric": "No disclosed data available."}]
        for rr, rec in enumerate(rows_in, start=data_start_row):
            fill = neutral_fill if (rr - data_start_row) % 2 == 0 else alt_fill
            metric_txt = str(rec.get("Metric") or "")
            for cc, header in enumerate(headers_in, start=1):
                value = rec.get(header)
                value = None if pd.isna(value) else _safe_cell(value)
                cell = ws.cell(row=rr, column=cc, value=value)
                cell.fill = copy(fill)
                cell.font = copy(body_font_local)
                cell.border = copy(thin_border_local)
                cell.alignment = Alignment(horizontal="left" if cc == 1 else "right", vertical="center", wrap_text=cc == 1)
                if cc > 1:
                    if metric_txt.endswith("(m)"):
                        cell.number_format = '#,##0.0'
                    else:
                        cell.number_format = '$0.000;($0.000);-'
            ws.row_dimensions[rr].height = 22.0
        return data_start_row + len(rows_in) + 1

    diagnostic_start_row = note_row + 5
    diagnostic_end_col = 11
    row_num = _write_diagnostic_section_title(
        diagnostic_start_row,
        "Derivative / Hedge Margin Diagnostics",
        (
            "Diagnostic translation of derivative P&L, OCI, AOCI and open hedge exposure into margin context. "
            "P&L rows affect reported margins; OCI/AOCI and net derivative balances are exposure or deferred items, not current-quarter margin impact."
        ),
        diagnostic_end_col,
    )
    quarterly_headers = ["Metric"] + quarter_labels_in_order
    row_num = _write_quarterly_impact_table(
        row_num,
        "Quarterly derivative impact on reported margin ($/gal)",
        quarterly_headers,
        quarterly_impact_rows,
    )
    quarterly_note_row = row_num
    ws.merge_cells(start_row=quarterly_note_row, start_column=1, end_row=quarterly_note_row, end_column=max(2, len(quarterly_headers)))
    quarterly_note = ws.cell(
        row=quarterly_note_row,
        column=1,
        value=(
            "Note: This table translates disclosed derivative/hedge income-statement impact into a reported-margin-equivalent $/gal. "
            "It is not a pure spot crush margin and does not include OCI/AOCI amounts that have not entered P&L. "
            "$/gal uses ethanol gallons produced as the diagnostic denominator. This is a reported-margin-equivalent scaling metric, not a contract-level hedge attribution."
        ),
    )
    quarterly_note.fill = copy(section_fill_local)
    quarterly_note.font = copy(body_font_local)
    quarterly_note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for cc in range(1, max(2, len(quarterly_headers)) + 1):
        ws.cell(row=quarterly_note_row, column=cc).fill = copy(section_fill_local)
        ws.cell(row=quarterly_note_row, column=cc).border = copy(thin_border_local)
    ws.row_dimensions[quarterly_note_row].height = 38.0
    row_num = quarterly_note_row + 2
    row_num = _write_diagnostic_table(
        row_num,
        "Derivative P&L per Gallon",
        ["Quarter", "Metric", "Amount ($m)", "Denominator", "Gallons (m)", "$/gal", "P&L line", "Interpretation"],
        pnl_per_gallon_rows,
        "DerivativePnlPerGallon",
    )
    row_num = _write_diagnostic_table(
        row_num,
        "Deferred / Balance-Sheet Hedge Exposure per Gallon",
        ["Quarter", "Metric", "Amount ($m)", "Denominator", "Gallons (m)", "$/gal", "Accounting status", "Interpretation"],
        deferred_rows,
        "DeferredHedgeExposurePerGallon",
    )
    row_num = _write_diagnostic_table(
        row_num,
        "Hedge Exposure by Margin Bucket",
        ["Quarter", "Commodity", "Instrument", "Accounting bucket", "Direction", "Net notional", "Unit", "Scale", "Likely P&L line", "Margin bucket", "Interpretation"],
        margin_bucket_rows,
        "HedgeExposureByMarginBucket",
    )
    for letter, width in {
        "A": 14,
        "B": 34,
        "C": 16,
        "D": 28,
        "E": 16,
        "F": 14,
        "G": 24,
        "H": 64,
        "I": 22,
        "J": 24,
        "K": 64,
    }.items():
        ws.column_dimensions[letter].width = max(float(ws.column_dimensions[letter].width or 0), float(width))

    diagnostics_note_row = row_num
    ws.merge_cells(start_row=diagnostics_note_row, start_column=1, end_row=diagnostics_note_row, end_column=diagnostic_end_col)
    diagnostics_note = ws.cell(
        row=diagnostics_note_row,
        column=1,
        value=(
            "Note: P&L per gallon metrics show reported income-statement impact during the quarter. OCI, AOCI and net derivative "
            "asset/liability per gallon are diagnostic exposure metrics, not current-quarter margin impact. Commodity/margin-bucket "
            "classifications are based on disclosure category and likely economic role; fair value and P&L by commodity are not "
            "disclosed unless explicitly shown by the company."
        ),
    )
    diagnostics_note.fill = copy(section_fill_local)
    diagnostics_note.font = copy(body_font_local)
    diagnostics_note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for cc in range(1, diagnostic_end_col + 1):
        ws.cell(row=diagnostics_note_row, column=cc).fill = copy(section_fill_local)
        ws.cell(row=diagnostics_note_row, column=cc).border = copy(thin_border_local)
    ws.row_dimensions[diagnostics_note_row].height = 42.0
    scale_note_row = diagnostics_note_row + 1
    ws.merge_cells(start_row=scale_note_row, start_column=1, end_row=scale_note_row, end_column=diagnostic_end_col)
    scale_note = ws.cell(
        row=scale_note_row,
        column=1,
        value="Note: Notional exposure is preserved on the company's disclosure scale. For GPRE, notional amounts are disclosed in thousands of units.",
    )
    scale_note.fill = copy(section_fill_local)
    scale_note.font = copy(body_font_local)
    scale_note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for cc in range(1, diagnostic_end_col + 1):
        ws.cell(row=scale_note_row, column=cc).fill = copy(section_fill_local)
        ws.cell(row=scale_note_row, column=cc).border = copy(thin_border_local)
    ws.row_dimensions[scale_note_row].height = 30.0


def write_derivative_crush_tests_sheet(
    deps: DerivativeOciBridgeRenderDeps,
    tables: Mapping[str, Any],
) -> None:
    """Write the GPRE-only derivative/crush diagnostic workbook surface.

    This sheet compares market/proxy crush lenses with reported margin after
    adding derivative P&L. It intentionally keeps OCI/AOCI and net derivative
    balances in lead/lag or exposure tables instead of treating them as
    current-period P&L.
    """
    runtime = deps.runtime
    wb = runtime["wb"]
    _safe_cell = runtime["_safe_cell"]
    _get_analysis_sheet_style_bundle = runtime["_get_analysis_sheet_style_bundle"]
    font_size = runtime["font_size"]
    header_size = runtime["header_size"]
    sheet_name = "Derivative_Crush_Tests"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    theme = _get_analysis_sheet_style_bundle()
    title_fill_local = copy(theme["title_fill"])
    section_fill_local = copy(theme["section_fill"])
    header_fill_local = copy(theme["header_fill"])
    neutral_fill = copy(theme["neutral_fill_alt"])
    alt_fill = copy(theme["neutral_fill"])
    thin_border_local = copy(theme["thin_border"])
    title_font_local = Font(bold=True, size=header_size + 2, color="FFFFFF")
    subtitle_font = Font(size=font_size, italic=True, color=str(theme["text_muted"]))
    header_font_local = Font(bold=True, size=header_size, color=str(theme["text_dark"]))
    body_font_local = Font(size=font_size, color=str(theme["text_dark"]))
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "A7"

    def _safe_table_value(value: Any) -> Any:
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        return _safe_cell(value)

    def _table_name(title: str) -> str:
        return re.sub(r"[^A-Za-z0-9]+", "", title)[:220] or "DerivativeCrushTable"

    def _write_merged_row(row_num: int, text: str, *, end_col: int, fill: PatternFill, font: Font, height: float) -> int:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=end_col)
        cell = ws.cell(row=row_num, column=1, value=text)
        cell.fill = copy(fill)
        cell.font = copy(font)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for cc in range(1, end_col + 1):
            ws.cell(row=row_num, column=cc).fill = copy(fill)
            ws.cell(row=row_num, column=cc).border = copy(thin_border_local)
        ws.row_dimensions[row_num].height = height
        return row_num + 1

    def _write_table(row_num: int, title: str, df_in: Any, *, note: str = "", table_name: Optional[str] = None) -> int:
        df = df_in.copy() if isinstance(df_in, pd.DataFrame) else pd.DataFrame()
        if df.empty:
            df = pd.DataFrame([{"Status": "No disclosed data available."}])
        headers = [str(col) for col in df.columns]
        end_col = max(1, len(headers))
        row_num = _write_merged_row(
            row_num,
            title,
            end_col=end_col,
            fill=section_fill_local,
            font=header_font_local,
            height=24.0,
        )
        if note:
            row_num = _write_merged_row(
                row_num,
                note,
                end_col=end_col,
                fill=neutral_fill,
                font=subtitle_font,
                height=34.0,
            )
        header_row = row_num
        for cc, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=cc, value=header)
            cell.fill = copy(header_fill_local)
            cell.font = copy(header_font_local)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = copy(thin_border_local)
        ws.row_dimensions[header_row].height = 28.0
        data_start = header_row + 1
        for rr, (_, rec) in enumerate(df.iterrows(), start=data_start):
            fill = neutral_fill if (rr - data_start) % 2 == 0 else alt_fill
            for cc, header in enumerate(headers, start=1):
                value = _safe_table_value(rec.get(header))
                cell = ws.cell(row=rr, column=cc, value=value)
                cell.fill = copy(fill)
                cell.font = copy(body_font_local)
                cell.border = copy(thin_border_local)
                cell.alignment = Alignment(
                    horizontal=(
                        "right"
                        if any(token in header for token in (" / gal", "MAE", "RMSE", "error", "improvement", "ratio", "Correlation", "slope", "R^2", "notional", "quarters", "Alpha", "Beta", "Gamma"))
                        else "left"
                    ),
                    vertical="center",
                    wrap_text=header in {"Formula", "Interpretation", "Notes / flags", "Possible explanation", "Data quality flag"},
                )
                low_header = header.lower()
                if header == "Quarter":
                    cell.number_format = "yyyy-mm-dd"
                elif "/ gal" in header or "$/gal" in header or "margin" in low_header or "error" in low_header or "improvement" in low_header or "slope" in low_header or header in {"Alpha", "Beta on proxy", "Gamma on derivative P&L"}:
                    cell.number_format = '$0.000;($0.000);-'
                elif header in {"Valid quarters", "Valid observations"}:
                    cell.number_format = '#,##0'
                elif header in {"Directional hit rate", "Coverage ratio"} and isinstance(value, (int, float)):
                    cell.number_format = "0.0%"
                elif header in {"Correlation", "R^2"}:
                    cell.number_format = "0.000"
                elif "notional" in low_header:
                    cell.number_format = '#,##0.0;(#,##0.0);-'
                elif header == "Gallons (m)":
                    cell.number_format = '#,##0.0'
            ws.row_dimensions[rr].height = 22.0
        data_end = data_start + len(df) - 1
        try:
            ref = f"A{header_row}:{get_column_letter(end_col)}{data_end}"
            table = Table(displayName=table_name or _table_name(title), ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(table)
        except Exception:
            pass

        header_index = {header: idx + 1 for idx, header in enumerate(headers)}
        for cf_header in ("Error improvement / gal", "Improvement vs Model A"):
            col_idx = header_index.get(cf_header)
            if col_idx:
                letter = get_column_letter(col_idx)
                rng = f"{letter}{data_start}:{letter}{data_end}"
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="E2F0D9")),
                )
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FCE4D6")),
                )
        residual_col = header_index.get("Residual after derivative adjustment / gal")
        if residual_col:
            letter = get_column_letter(residual_col)
            rng = f"{letter}{data_start}:{letter}{data_end}"
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="greaterThan", formula=["0.03"], fill=PatternFill("solid", fgColor="FCE4D6")),
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="lessThan", formula=["-0.03"], fill=PatternFill("solid", fgColor="FCE4D6")),
            )
        return data_end + 2

    max_col = 10
    row_idx = 1
    row_idx = _write_merged_row(
        row_idx,
        "Derivative_Crush_Tests",
        end_col=max_col,
        fill=title_fill_local,
        font=title_font_local,
        height=26.0,
    )
    row_idx = _write_merged_row(
        row_idx,
        (
            "This sheet tests whether derivative/hedge disclosures improve explanation of reported margin versus "
            "the market/proxy crush model. Current-quarter tests use derivative P&L only. OCI, AOCI and net "
            "derivative balances are treated as deferred or balance-sheet exposure signals, not current-quarter P&L."
        ),
        end_col=max_col,
        fill=section_fill_local,
        font=subtitle_font,
        height=38.0,
    )
    row_idx = _write_merged_row(
        row_idx,
        (
            "Derivative P&L is a reported-margin adjustment, not a pure spot crush margin. Negative derivative "
            "P&L may reflect hedging/timing/risk management, not necessarily poor execution."
        ),
        end_col=max_col,
        fill=neutral_fill,
        font=subtitle_font,
        height=30.0,
    )
    row_idx += 1
    row_idx = _write_table(
        row_idx,
        "Key Diagnostic Takeaways",
        tables.get("key_takeaways"),
        note="Conservative read of the diagnostics below; recommendations do not alter production model outputs.",
        table_name="DerivativeCrushKeyTakeaways",
    )
    row_idx = _write_table(
        row_idx,
        "Model Accuracy Summary",
        tables.get("model_summary"),
        table_name="DerivativeCrushModelAccuracy",
    )
    row_idx = _write_table(
        row_idx,
        "Q4 / Quarterization Sensitivity",
        tables.get("q4_quarterization_sensitivity"),
        note="Checks whether Q4 annual-minus-Q1-Q3 quarterization distorts model accuracy. Diagnostic only.",
        table_name="DerivativeCrushQ4Sensitivity",
    )
    row_idx = _write_table(
        row_idx,
        "Ex-Derivative Margin Test",
        tables.get("ex_derivative_margin_test"),
        note="Physical-margin diagnostic: compares proxy lenses to reported margin before and after removing derivative P&L.",
        table_name="DerivativeCrushExDerivativeMargin",
    )
    row_idx = _write_table(
        row_idx,
        "Clean Margin Bridge Diagnostic",
        tables.get("clean_margin_bridge"),
        note="Diagnostic clean margin subtracts only available explicit bridge items; missing components remain flagged, not estimated.",
        table_name="DerivativeCrushCleanMarginBridge",
    )
    row_idx = _write_table(
        row_idx,
        "Target-Specific Model Accuracy",
        tables.get("target_specific_model_accuracy"),
        table_name="DerivativeCrushTargetAccuracy",
    )
    row_idx = _write_table(
        row_idx,
        "Revenue / COGS Side Decomposition",
        tables.get("revenue_cogs_decomposition"),
        note="Tests whether revenue-side and COGS-side derivative components explain reported-margin error differently.",
        table_name="DerivativeCrushRevenueCogsDecomp",
    )
    row_idx = _write_table(
        row_idx,
        "Volume / Utilization Residual Test",
        tables.get("volume_utilization_summary"),
        note="Summary relationships between residuals, gallons, utilization, and production timing. Diagnostic only.",
        table_name="DerivativeCrushVolumeUtilizationSummary",
    )
    row_idx = _write_table(
        row_idx,
        "Volume / Utilization Quarter Detail",
        tables.get("volume_utilization_quarterly"),
        table_name="DerivativeCrushVolumeUtilizationDetail",
    )
    row_idx = _write_table(
        row_idx,
        "Basis and Energy Residual Screen",
        tables.get("basis_energy_summary"),
        note="Screens local basis, natural gas, and coproduct proxies against residuals. Production proxy unchanged.",
        table_name="DerivativeCrushBasisEnergySummary",
    )
    row_idx = _write_table(
        row_idx,
        "Basis and Energy Quarter Detail",
        tables.get("basis_energy_quarterly"),
        table_name="DerivativeCrushBasisEnergyDetail",
    )
    row_idx = _write_table(
        row_idx,
        "AOCI Future Reclass Tracker",
        tables.get("aoci_future_reclass_summary"),
        note="AOCI/OCI/net derivative balances are lead variables only; they are not current-quarter P&L.",
        table_name="DerivativeCrushAociFutureSummary",
    )
    row_idx = _write_table(
        row_idx,
        "AOCI Future Reclass Quarter Detail",
        tables.get("aoci_future_reclass_tracker"),
        table_name="DerivativeCrushAociFutureDetail",
    )
    row_idx = _write_table(
        row_idx,
        "Reported Margin Reconciliation: Market Proxy vs Derivative-Adjusted",
        tables.get("reconciliation"),
        table_name="DerivativeCrushReconciliation",
    )
    row_idx = _write_table(
        row_idx,
        "Coefficient Diagnostic",
        tables.get("coefficient_diagnostic"),
        note="Regression diagnostics are not production model coefficients and are not promoted automatically.",
        table_name="DerivativeCrushCoefficientDiagnostic",
    )
    row_idx = _write_table(
        row_idx,
        "Lagged Derivative P&L Tests",
        tables.get("lagged_derivative_pnl_tests"),
        note="Timing variants test recognition mismatch; OCI/AOCI and balance-sheet exposure are excluded from current-quarter P&L.",
        table_name="DerivativeCrushLaggedDerivativePnl",
    )
    row_idx = _write_table(
        row_idx,
        "Quarterly derivative impact on reported margin ($/gal)",
        tables.get("quarterly_derivative_impact"),
        note=(
            "$/gal uses ethanol gallons produced as the diagnostic denominator where available. "
            "This is a reported-margin-equivalent scaling metric, not a contract-level hedge attribution."
        ),
        table_name="DerivativeCrushQuarterlyImpact",
    )
    row_idx = _write_table(
        row_idx,
        "Lead/Lag Tests: Deferred Hedge Balances vs Future P&L",
        tables.get("lead_lag_summary"),
        note="OCI/AOCI/net derivative balances are lead variables only; they are not current-quarter P&L.",
        table_name="DerivativeCrushLeadLagSummary",
    )
    row_idx = _write_table(
        row_idx,
        "Lead/Lag Detail",
        tables.get("lead_lag_detail"),
        table_name="DerivativeCrushLeadLagDetail",
    )
    row_idx = _write_table(
        row_idx,
        "Residual Driver Screen",
        tables.get("residual_driver_screen"),
        note="Screens remaining gaps after market/proxy crush and derivative P&L; correlations are diagnostic only.",
        table_name="DerivativeCrushResidualDriverScreen",
    )
    row_idx = _write_table(
        row_idx,
        "Hedge Slippage Diagnostic",
        tables.get("slippage"),
        note="Threshold is $0.03/gal. Flags are risk-management diagnostics, not execution judgments.",
        table_name="DerivativeCrushSlippage",
    )
    row_idx = _write_table(
        row_idx,
        "Open Hedge Exposure by Margin Bucket",
        tables.get("exposure_buckets"),
        note="Coverage ratios are intentionally marked not available until compatible physical-volume denominators exist.",
        table_name="DerivativeCrushExposureBuckets",
    )
    row_idx = _write_table(
        row_idx,
        "Residual Analysis After Derivative Adjustment",
        tables.get("residual"),
        table_name="DerivativeCrushResidual",
    )
    for letter, width in {
        "A": 22,
        "B": 18,
        "C": 20,
        "D": 22,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 48,
        "J": 52,
        "K": 22,
        "L": 24,
    }.items():
        ws.column_dimensions[letter].width = max(float(ws.column_dimensions[letter].width or 0), float(width))
