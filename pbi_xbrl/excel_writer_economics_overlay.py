from __future__ import annotations

import time
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Any, Callable, Dict, Optional

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class GpreOverlaySupportInputs:
    wb: Any
    ws: Any
    row_idx: int
    is_gpre_profile: bool
    has_gpre_commercial_setup: bool
    model_result: Dict[str, Any]
    proxy_implied_results_bundle: Dict[str, Any]
    bridge_panel_rows: Dict[str, int]
    ticker_root: Optional[Path]
    current_ref: Dict[str, str]
    thesis_ref: Dict[str, str]
    title_fill: Any
    title_font: Any
    header_fill: Any
    bold_font: Any
    body_font: Any
    thin_border: Any
    align_center: Any
    align_center_wrap: Any
    align_left_center_wrap: Any
    zebra_fill_light: Any
    sandbox_writer: Callable[[Any, int, Dict[str, Any]], Dict[str, Any]]
    add_comment: Callable[[str, Any], None]
    gpre_fitted_live_formula: Callable[[str, str], Optional[str]]
    gpre_formula_note: Callable[[str], str]
    gpre_preview_frame_value: Callable[[str, str], Optional[float]]
    gpre_preview_frame_note: Callable[[str, str], str]
    gpre_model_live_formula: Callable[[str, str, str], Optional[str]]
    gpre_model_formula_note: Callable[[str, str], str]
    gpre_model_preview_frame_value: Callable[[str, str], Optional[float]]
    gpre_model_preview_frame_note: Callable[[str, str], str]
    record_writer_substage: Callable[[str, float], None]


@dataclass(frozen=True)
class GpreOverlaySupportResult:
    sandbox_layout: Dict[str, Any] = field(default_factory=dict)
    proxy_comp_end_row: int = 0
    proxy_comp_title_row: int = 0
    proxy_comp_header_row: int = 0
    official_proxy_comp_row: int = 0
    fitted_proxy_comp_row: int = 0
    best_forward_proxy_comp_row: int = 0


def write_gpre_basis_proxy_sidecars(
    ticker_root: Optional[Path],
    model_result: Dict[str, Any],
) -> None:
    if not isinstance(model_result, dict):
        return
    quarterly_df = model_result.get("quarterly_df")
    if not isinstance(quarterly_df, pd.DataFrame) or quarterly_df.empty:
        return
    if ticker_root is None:
        return
    out_dir = ticker_root / "basis_proxy"
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
        quarterly_df.to_csv(out_dir / "gpre_basis_proxy_quarterly.csv", index=False)
        quarterly_df.to_parquet(out_dir / "gpre_basis_proxy_quarterly.parquet", index=False)
        (out_dir / "gpre_basis_proxy_summary.md").write_text(
            str(model_result.get("summary_markdown") or "").strip() + "\n",
            encoding="utf-8",
        )
    except Exception:
        pass


def _overlay_model_label(model_key: Any) -> str:
    key_txt = str(model_key or "").strip()
    if not key_txt:
        return "n/a"
    explicit = {
        "process_current_quarter_avg": "Process current avg",
        "process_front_loaded": "Process front-loaded",
        "process_quarter_open_blend_exec_penalty": "Process q-open + severe ops penalty",
        "process_utilization_regime_residual": "Process utilization regime residual",
        "process_market_process_ensemble_35_65": "Market/process ensemble 35/65",
        "process_locked_share_asymmetric_passthrough": "Locked-share asymmetric passthrough",
        "process_prior_gap_carryover_small": "Prior-gap carryover small",
        "process_prior_disturbance_carryover": "Prior-disturbance carryover",
    }
    if key_txt in explicit:
        return explicit[key_txt]
    return key_txt.replace("_", " ").replace("gpre", "GPRE").title()


def _compact_proxy_comparison_comment(note_txt: Any, model_key: Any, frame_key: Any) -> str:
    note = " ".join(str(note_txt or "").replace("\n", " ").split())
    phase_label_map = {
        "prior_quarter": "Prior quarter",
        "quarter_open": "Quarter-open",
        "current_qtd": "Current QTD",
        "next_quarter_thesis": "Next quarter",
    }
    phase_label = str(phase_label_map.get(str(frame_key or "").strip(), "Preview")).strip()
    model_label = _overlay_model_label(model_key)
    short_label_map = {
        "Process utilization regime residual": "Process util regime residual",
        "Process Utilization Regime Residual": "Process util regime residual",
        "Process Quarter Open Blend": "Process q-open blend",
        "Process q-open + severe ops penalty": "Process q-open ops penalty",
        "Process utilization regime blend": "Process util regime blend",
    }
    short_model_label = str(short_label_map.get(model_label, model_label)).strip()

    if note:
        for pattern in (
            r"^\s*Current fitted preview uses(?: the same)?\s+",
            r"^\s*Next-quarter fitted preview uses(?: the same)?\s+",
            r"^\s*Quarter-open fitted preview uses(?: the same)?\s+",
            r"^\s*Prior-quarter fitted preview uses(?: the same)?\s+",
            r"^\s*Current fitted preview keeps\s+",
            r"^\s*Next-quarter fitted preview keeps\s+",
            r"^\s*Quarter-open fitted preview keeps\s+",
            r"^\s*Prior-quarter fitted preview keeps\s+",
        ):
            note = re.sub(pattern, "", note, flags=re.I)
        replacements = {
            "on top of the incumbent ": "",
            "the incumbent ": "",
            "extra bounded ": "",
            "residual-drag": "residual drag",
            "lower-utilization": "low-utilization",
            "severe-execution": "severe execution",
            "process-execution": "process execution",
        }
        for old_txt, new_txt in replacements.items():
            note = note.replace(old_txt, new_txt)
        note = " ".join(note.strip(" .;:-").split())
        if note:
            note = note.rstrip(".") + "."
            if len(note.split()) <= 12:
                return note

    fallback = f"{phase_label}: {short_model_label}."
    if len(fallback.split()) <= 12:
        return fallback
    return f"{phase_label} preview."


def _write_proxy_implied_results_panel(
    inputs: GpreOverlaySupportInputs,
    *,
    official_proxy_comp_row: int,
    fitted_proxy_comp_row: int,
    best_forward_proxy_comp_row: int = 0,
) -> None:
    ws = inputs.ws
    panel_title_row = int(inputs.bridge_panel_rows.get("panel_title_row") or 0)
    panel_header_row = int(inputs.bridge_panel_rows.get("panel_header_row") or 0)
    approx_bridge_row = int(inputs.bridge_panel_rows.get("approx_market_crush_proxy") or 0)
    fitted_bridge_row = int(inputs.bridge_panel_rows.get("gpre_crush_proxy") or 0)
    forward_bridge_row = int(inputs.bridge_panel_rows.get("best_forward_lens_proxy") or 0)
    if not all(
        row_val > 0
        for row_val in (
            panel_title_row,
            panel_header_row,
            approx_bridge_row,
            fitted_bridge_row,
        )
    ):
        return

    helper_gallons_row = int(inputs.bridge_panel_rows.get("underlying_crush_margin") or 0)
    helper_basis_row = int(inputs.bridge_panel_rows.get("reported_consolidated_crush_margin") or 0)
    if helper_gallons_row <= 0 or helper_basis_row <= 0:
        helper_gallons_row = fitted_bridge_row + 2
        helper_basis_row = fitted_bridge_row + 3

    panel_start_col = 14
    panel_end_col = 21
    helper_label_start_col = 22
    helper_label_end_col = 24
    frame_spans = {
        "prior_quarter": (14, 15),
        "quarter_open": (16, 17),
        "current_qtd": (18, 19),
        "next_quarter_thesis": (20, 21),
    }
    proxy_source_cols = {
        "prior_quarter": 2,
        "quarter_open": 4,
        "current_qtd": 6,
        "next_quarter_thesis": 8,
    }
    proxy_implied_frames = dict(inputs.proxy_implied_results_bundle.get("frames") or {})
    neighbor_width = ws.column_dimensions["U"].width or ws.column_dimensions["T"].width or 13.86
    for col_idx in range(helper_label_start_col, helper_label_end_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = neighbor_width

    ws.merge_cells(
        start_row=panel_title_row,
        start_column=panel_start_col,
        end_row=panel_title_row,
        end_column=panel_end_col,
    )
    panel_title_cell = ws.cell(
        row=panel_title_row,
        column=panel_start_col,
        value=str(inputs.proxy_implied_results_bundle.get("title") or "Proxy-implied results ($m)"),
    )
    panel_title_cell.fill = inputs.title_fill
    panel_title_cell.font = inputs.title_font
    panel_title_cell.border = inputs.thin_border
    panel_title_cell.alignment = inputs.align_center_wrap
    for cc in range(panel_start_col, panel_end_col + 1):
        ws.cell(row=panel_title_row, column=cc).fill = inputs.title_fill
        ws.cell(row=panel_title_row, column=cc).font = inputs.title_font
        ws.cell(row=panel_title_row, column=cc).border = inputs.thin_border
        ws.cell(row=panel_title_row, column=cc).alignment = inputs.align_center_wrap
    ws.row_dimensions[panel_title_row].height = 18.0
    panel_note = str(inputs.proxy_implied_results_bundle.get("note") or "").strip()
    if panel_note:
        inputs.add_comment(f"{get_column_letter(panel_start_col)}{panel_title_row}", panel_note)

    for cc in range(panel_start_col, panel_end_col + 1):
        ws.cell(row=panel_header_row, column=cc).fill = inputs.header_fill
        ws.cell(row=panel_header_row, column=cc).font = inputs.bold_font
        ws.cell(row=panel_header_row, column=cc).border = inputs.thin_border
        ws.cell(row=panel_header_row, column=cc).alignment = inputs.align_center_wrap
    for frame_key, (start_col, end_col) in frame_spans.items():
        ws.merge_cells(
            start_row=panel_header_row,
            start_column=start_col,
            end_row=panel_header_row,
            end_column=end_col,
        )
        ws.cell(
            row=panel_header_row,
            column=start_col,
            value=str((proxy_implied_frames.get(frame_key) or {}).get("frame_label") or ""),
        )

    panel_rows = [approx_bridge_row, fitted_bridge_row]
    if forward_bridge_row > 0:
        panel_rows.append(forward_bridge_row)
    panel_rows.extend([helper_gallons_row, helper_basis_row])

    for row_num_local in panel_rows:
        fill_ref = copy(inputs.zebra_fill_light)
        for cc in range(panel_start_col, panel_end_col + 1):
            ws.cell(row=row_num_local, column=cc).fill = fill_ref
            ws.cell(row=row_num_local, column=cc).font = inputs.body_font
            ws.cell(row=row_num_local, column=cc).border = inputs.thin_border
            ws.cell(row=row_num_local, column=cc).alignment = inputs.align_center_wrap
    for row_num_local in panel_rows:
        for start_col, end_col in frame_spans.values():
            ws.merge_cells(
                start_row=row_num_local,
                start_column=start_col,
                end_row=row_num_local,
                end_column=end_col,
            )

    for title_row, title_txt in (
        (helper_gallons_row, "Implied gallons assumption"),
        (helper_basis_row, "Volume basis"),
    ):
        ws.merge_cells(
            start_row=title_row,
            start_column=helper_label_start_col,
            end_row=title_row,
            end_column=helper_label_end_col,
        )
        for cc in range(helper_label_start_col, helper_label_end_col + 1):
            ws.cell(row=title_row, column=cc).fill = inputs.header_fill
            ws.cell(row=title_row, column=cc).font = inputs.bold_font
            ws.cell(row=title_row, column=cc).border = inputs.thin_border
            ws.cell(row=title_row, column=cc).alignment = inputs.align_left_center_wrap
        ws.cell(row=title_row, column=helper_label_start_col, value=title_txt)

    for frame_key, (start_col, _end_col) in frame_spans.items():
        frame_rec = dict(proxy_implied_frames.get(frame_key) or {})
        proxy_src_col = int(proxy_source_cols.get(frame_key) or 0)
        gallons_cell_ref = f"{get_column_letter(start_col)}{helper_gallons_row}"
        official_proxy_ref = f"{get_column_letter(proxy_src_col)}{official_proxy_comp_row}"
        fitted_proxy_ref = f"{get_column_letter(proxy_src_col)}{fitted_proxy_comp_row}"
        ws.cell(
            row=approx_bridge_row,
            column=start_col,
            value=f'=IF(AND(ISNUMBER({official_proxy_ref}),ISNUMBER({gallons_cell_ref})),{official_proxy_ref}*{gallons_cell_ref},"")',
        ).number_format = "0.0;-0.0"
        ws.cell(
            row=fitted_bridge_row,
            column=start_col,
            value=f'=IF(AND(ISNUMBER({fitted_proxy_ref}),ISNUMBER({gallons_cell_ref})),{fitted_proxy_ref}*{gallons_cell_ref},"")',
        ).number_format = "0.0;-0.0"
        if forward_bridge_row > 0 and best_forward_proxy_comp_row > 0:
            forward_proxy_ref = f"{get_column_letter(proxy_src_col)}{best_forward_proxy_comp_row}"
            ws.cell(
                row=forward_bridge_row,
                column=start_col,
                value=f'=IF(AND(ISNUMBER({forward_proxy_ref}),ISNUMBER({gallons_cell_ref})),{forward_proxy_ref}*{gallons_cell_ref},"")',
            ).number_format = "0.0;-0.0"

        implied_gallons_display_num = pd.to_numeric(frame_rec.get("implied_gallons_million_display"), errors="coerce")
        gallons_cell = ws.cell(row=helper_gallons_row, column=start_col)
        if pd.notna(implied_gallons_display_num):
            gallons_cell.value = float(implied_gallons_display_num)
            gallons_cell.number_format = '0.0"m gal"'
        else:
            gallons_cell.value = ""
        basis_cell = ws.cell(row=helper_basis_row, column=start_col, value=str(frame_rec.get("volume_basis_display") or "Unavailable"))
        basis_cell.alignment = inputs.align_center_wrap
        note_bits = [
            str(frame_rec.get("volume_basis_comment") or "").strip(),
            str(frame_rec.get("reason_unavailable") or "").strip(),
        ]
        note_txt = " ".join(bit for bit in note_bits if bit)
        if note_txt:
            inputs.add_comment(f"{get_column_letter(start_col)}{helper_gallons_row}", note_txt)


def write_gpre_basis_proxy_overlay_support(
    inputs: GpreOverlaySupportInputs,
) -> GpreOverlaySupportResult:
    overlay_sandbox_started = time.perf_counter()
    if not (inputs.is_gpre_profile and inputs.has_gpre_commercial_setup):
        inputs.record_writer_substage(
            "write_excel.drivers.render.economics_overlay.basis_proxy_sandbox",
            overlay_sandbox_started,
        )
        return GpreOverlaySupportResult(proxy_comp_end_row=inputs.row_idx - 1)
    write_gpre_basis_proxy_sidecars(inputs.ticker_root, inputs.model_result)
    sandbox_ws = inputs.wb.create_sheet("Basis_Proxy_Sandbox")
    sandbox_layout = inputs.sandbox_writer(sandbox_ws, 1, inputs.model_result)
    sandbox_process_margin_refs = (
        ((sandbox_layout.get("approx_market_crush_build_up") or {}).get("process_margin_refs"))
        if isinstance(sandbox_layout, dict)
        else {}
    ) or {}
    inputs.record_writer_substage(
        "write_excel.drivers.render.economics_overlay.basis_proxy_sandbox",
        overlay_sandbox_started,
    )

    overlay_proxy_comparison_started = time.perf_counter()
    ws = inputs.ws
    proxy_comp_title_row = inputs.row_idx
    proxy_comp_note_row = proxy_comp_title_row + 1
    proxy_comp_header_row = proxy_comp_title_row + 2
    official_proxy_comp_row = proxy_comp_header_row + 1
    fitted_proxy_comp_row = proxy_comp_header_row + 2
    best_forward_proxy_comp_row = proxy_comp_header_row + 3
    proxy_comp_end_row = best_forward_proxy_comp_row
    production_winner_model_key = str(
        inputs.model_result.get("production_winner_model_key")
        or inputs.model_result.get("gpre_proxy_model_key")
        or ""
    )
    best_forward_lens_model_key = str(inputs.model_result.get("best_forward_lens_model_key") or "")
    ws.merge_cells(start_row=proxy_comp_title_row, start_column=1, end_row=proxy_comp_title_row, end_column=9)
    proxy_title_cell = ws.cell(row=proxy_comp_title_row, column=1, value="Proxy comparison ($/gal)")
    proxy_title_cell.fill = inputs.title_fill
    proxy_title_cell.font = inputs.title_font
    proxy_title_cell.alignment = inputs.align_center
    for cc in range(1, 10):
        ws.cell(row=proxy_comp_title_row, column=cc).fill = inputs.title_fill
        ws.cell(row=proxy_comp_title_row, column=cc).font = inputs.title_font
        ws.cell(row=proxy_comp_title_row, column=cc).border = inputs.thin_border
        ws.cell(row=proxy_comp_title_row, column=cc).alignment = inputs.align_center_wrap
    ws.row_dimensions[proxy_comp_title_row].height = 18.0
    ws.merge_cells(start_row=proxy_comp_note_row, start_column=1, end_row=proxy_comp_note_row, end_column=21)
    proxy_note_text = (
        "Official row = Approximate market crush | "
        "Fitted row = GPRE crush proxy | "
        f"Production winner = {_overlay_model_label(production_winner_model_key)} | "
        f"Best forward lens = {_overlay_model_label(best_forward_lens_model_key)}"
    )
    proxy_note_fill = PatternFill(fill_type="solid", fgColor="EDF4FA")
    proxy_note_cell = ws.cell(row=proxy_comp_note_row, column=1, value=proxy_note_text)
    proxy_note_cell.fill = copy(proxy_note_fill)
    proxy_note_cell.font = copy(inputs.body_font)
    proxy_note_cell.border = copy(inputs.thin_border)
    proxy_note_cell.alignment = inputs.align_left_center_wrap
    for cc in range(1, 22):
        ws.cell(row=proxy_comp_note_row, column=cc).fill = copy(proxy_note_fill)
        ws.cell(row=proxy_comp_note_row, column=cc).font = copy(inputs.body_font)
        ws.cell(row=proxy_comp_note_row, column=cc).border = copy(inputs.thin_border)
        ws.cell(row=proxy_comp_note_row, column=cc).alignment = inputs.align_left_center_wrap
    ws.row_dimensions[proxy_comp_note_row].height = 24.0
    proxy_header_spans = [
        (1, 1, "Proxy row"),
        (2, 3, "Prior quarter"),
        (4, 5, "Quarter-open proxy"),
        (6, 7, "Current QTD"),
        (8, 9, "Next quarter"),
    ]
    for start_col, end_col, hdr in proxy_header_spans:
        if end_col > start_col:
            ws.merge_cells(start_row=proxy_comp_header_row, start_column=start_col, end_row=proxy_comp_header_row, end_column=end_col)
        cell = ws.cell(row=proxy_comp_header_row, column=start_col, value=hdr)
        cell.fill = inputs.header_fill
        cell.font = inputs.bold_font
        cell.border = inputs.thin_border
        cell.alignment = inputs.align_center_wrap
        for cc in range(start_col, end_col + 1):
            ws.cell(row=proxy_comp_header_row, column=cc).fill = inputs.header_fill
            ws.cell(row=proxy_comp_header_row, column=cc).font = inputs.bold_font
            ws.cell(row=proxy_comp_header_row, column=cc).border = inputs.thin_border
            ws.cell(row=proxy_comp_header_row, column=cc).alignment = inputs.align_center_wrap
    ws.row_dimensions[proxy_comp_header_row].height = 21.0
    best_forward_model_key = best_forward_lens_model_key or production_winner_model_key
    proxy_table_rows = [
        {"row_num": official_proxy_comp_row, "label": "Approximate market crush ($/gal)", "frame_group": "official_frames", "model_key": ""},
        {"row_num": fitted_proxy_comp_row, "label": "GPRE crush proxy ($/gal)", "frame_group": "gpre_proxy_frames", "model_key": production_winner_model_key},
        {"row_num": best_forward_proxy_comp_row, "label": "Best forward lens ($/gal)", "frame_group": "gpre_proxy_frames", "model_key": best_forward_model_key},
    ]
    frame_order = ("prior_quarter", "quarter_open", "current_qtd", "next_quarter_thesis")
    official_proxy_refs = {
        frame_key: (
            f'=IF(ISNUMBER({sandbox_process_margin_refs[frame_key]}),{sandbox_process_margin_refs[frame_key]},"")'
            if str(sandbox_process_margin_refs.get(frame_key) or "").strip()
            else '=""'
        )
        for frame_key in frame_order
    }
    for row_cfg in proxy_table_rows:
        row_num = int(row_cfg.get("row_num") or 0)
        label_txt = str(row_cfg.get("label") or "").strip()
        frame_group = str(row_cfg.get("frame_group") or "").strip()
        model_key = str(row_cfg.get("model_key") or "").strip()
        ws.cell(row=row_num, column=1, value=label_txt)
        value_spans = [(2, 3), (4, 5), (6, 7), (8, 9)]
        for start_col, end_col in value_spans:
            ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
        fill_ref = copy(inputs.zebra_fill_light)
        for cc in range(1, 10):
            cell = ws.cell(row=row_num, column=cc)
            cell.fill = fill_ref
            cell.font = inputs.body_font
            cell.border = inputs.thin_border
            cell.alignment = inputs.align_left_center_wrap if cc == 1 else inputs.align_center
        for (start_col, _end_col), frame_key in zip(value_spans, frame_order):
            cell = ws.cell(row=row_num, column=start_col)
            if frame_group == "official_frames":
                cell.value = official_proxy_refs.get(frame_key, "")
                cell.number_format = "#,##0.000"
                continue
            if frame_key in {"current_qtd", "next_quarter_thesis"}:
                live_ref = inputs.gpre_model_live_formula(
                    model_key,
                    frame_key,
                    inputs.current_ref.get("ethanol_price", "") if frame_key == "current_qtd" else inputs.thesis_ref.get("ethanol_price", ""),
                )
                if live_ref:
                    cell.value = live_ref
                    cell.number_format = "#,##0.000"
                    fitted_note = inputs.gpre_model_formula_note(model_key, frame_key)
                    if fitted_note:
                        inputs.add_comment(
                            f"{get_column_letter(start_col)}{row_num}",
                            _compact_proxy_comparison_comment(fitted_note, model_key, frame_key),
                        )
                    continue
            if model_key:
                value_num = inputs.gpre_model_preview_frame_value(model_key, frame_key)
            else:
                value_num = inputs.gpre_preview_frame_value(frame_group, frame_key)
            if value_num is not None:
                cell.value = value_num
                cell.number_format = "#,##0.000"
                fitted_note = (
                    inputs.gpre_model_preview_frame_note(model_key, frame_key)
                    if model_key
                    else inputs.gpre_preview_frame_note(frame_group, frame_key)
                )
                if fitted_note:
                    inputs.add_comment(
                        f"{get_column_letter(start_col)}{row_num}",
                        _compact_proxy_comparison_comment(fitted_note, model_key, frame_key),
                    )
        ws.row_dimensions[row_num].height = 19.5
    inputs.record_writer_substage(
        "write_excel.drivers.render.economics_overlay.proxy_comparison",
        overlay_proxy_comparison_started,
    )

    if inputs.bridge_panel_rows and inputs.proxy_implied_results_bundle:
        overlay_proxy_implied_started = time.perf_counter()
        _write_proxy_implied_results_panel(
            inputs,
            official_proxy_comp_row=official_proxy_comp_row,
            fitted_proxy_comp_row=fitted_proxy_comp_row,
            best_forward_proxy_comp_row=best_forward_proxy_comp_row,
        )
        inputs.record_writer_substage(
            "write_excel.drivers.render.economics_overlay.proxy_implied_results",
            overlay_proxy_implied_started,
        )

    return GpreOverlaySupportResult(
        sandbox_layout=sandbox_layout if isinstance(sandbox_layout, dict) else {},
        proxy_comp_end_row=proxy_comp_end_row,
        proxy_comp_title_row=proxy_comp_title_row,
        proxy_comp_header_row=proxy_comp_header_row,
        official_proxy_comp_row=official_proxy_comp_row,
        fitted_proxy_comp_row=fitted_proxy_comp_row,
        best_forward_proxy_comp_row=best_forward_proxy_comp_row,
    )
