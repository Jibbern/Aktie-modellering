"""Economics_Overlay bridge-to-reported row writer."""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date
from typing import Any, Callable, Dict, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class GpreEconomicsOverlayBridgeDeps:
    ws: Any
    row_idx: int
    is_gpre_profile: bool
    gpre_commercial_setup_rows: Sequence[Mapping[str, Any]]
    overlay_display_quarters: Sequence[Any]
    overlay_gpre_end_col: int
    row_map: Mapping[Tuple[str, Any], Mapping[str, Any]]
    bridge_bundle_map: Mapping[Any, Mapping[str, Any]]
    derivative_bridge_by_quarter: Mapping[Any, Mapping[str, Any]]
    bridge_templates: Sequence[Any]
    market_input_templates: Sequence[Any]
    gpre_basis_quarter_map: Mapping[Any, Mapping[str, Any]]
    current_overlay_model_key: str
    best_forward_overlay_model_key: str
    overlay_model_key_to_pred_col: Mapping[str, str]
    gpre_bridge_panel_rows: Mapping[str, int]
    gpre_reported_margin_by_quarter: Mapping[Any, float]
    gpre_denominator_policy_by_quarter: Mapping[Any, str]
    overlay_section_row_height: float
    overlay_intro_row_height: float
    overlay_header_row_height: float
    overlay_support_row_height: float
    header_fill: Any
    thin_border: Any
    bold_font: Any
    body_font: Any
    zebra_fill_light: Any
    write_section_bar: Callable[..., int]
    write_overlay_intro: Callable[..., int]
    add_comment: Callable[..., None]
    overlay_coefficient_detail: Callable[..., Any]
    pick_market_reference: Callable[..., Any]
    overlay_model_leaderboard_row: Callable[..., Any]
    overlay_model_label: Callable[..., str]
    driver_source_comment: Callable[..., Any]
    driver_source_note: Callable[..., str]


@dataclass(frozen=True)
class GpreEconomicsOverlayBridgeResult:
    row_idx: int
    bridge_separator_rows: list[int]
    gpre_bridge_panel_rows: dict[str, int]
    gpre_reported_margin_by_quarter: dict[Any, float]
    gpre_denominator_policy_by_quarter: dict[Any, str]


def _derivative_bridge_record(
    derivative_bridge_by_quarter: Mapping[Any, Mapping[str, Any]],
    qd_in: Any,
) -> Dict[str, Any]:
    if not isinstance(qd_in, date):
        return {}
    return dict(derivative_bridge_by_quarter.get(qd_in) or {})


def _derivative_usd_to_millions(usd_value: Any) -> Optional[float]:
    val = pd.to_numeric(usd_value, errors="coerce")
    if pd.isna(val):
        return None
    return float(val) / 1_000_000.0


def write_gpre_economics_overlay_bridge_to_reported_section(
    deps: GpreEconomicsOverlayBridgeDeps,
) -> GpreEconomicsOverlayBridgeResult:
    ws = deps.ws
    row_num = int(deps.row_idx)
    is_gpre_profile = bool(deps.is_gpre_profile)
    gpre_commercial_setup_rows = list(deps.gpre_commercial_setup_rows or [])
    overlay_display_quarters = list(deps.overlay_display_quarters or [])
    overlay_gpre_end_col = int(deps.overlay_gpre_end_col)
    row_map = deps.row_map
    bridge_bundle_map = deps.bridge_bundle_map
    derivative_bridge_by_quarter = deps.derivative_bridge_by_quarter
    bridge_templates = list(deps.bridge_templates or [])
    market_input_templates = list(deps.market_input_templates or [])
    gpre_basis_quarter_map = deps.gpre_basis_quarter_map
    current_overlay_model_key = str(deps.current_overlay_model_key or "").strip()
    best_forward_overlay_model_key = str(deps.best_forward_overlay_model_key or "").strip()
    overlay_model_key_to_pred_col = dict(deps.overlay_model_key_to_pred_col or {})
    gpre_bridge_panel_rows = dict(deps.gpre_bridge_panel_rows or {})
    gpre_reported_margin_by_quarter = dict(deps.gpre_reported_margin_by_quarter or {})
    gpre_denominator_policy_by_quarter = dict(deps.gpre_denominator_policy_by_quarter or {})
    overlay_section_row_height = deps.overlay_section_row_height
    overlay_intro_row_height = deps.overlay_intro_row_height
    overlay_header_row_height = deps.overlay_header_row_height
    overlay_support_row_height = deps.overlay_support_row_height
    header_fill = deps.header_fill
    thin_border = deps.thin_border
    bold_font = deps.bold_font
    body_font = deps.body_font
    zebra_fill_light = deps.zebra_fill_light
    _write_section_bar = deps.write_section_bar
    _write_overlay_intro = deps.write_overlay_intro
    _add_comment = deps.add_comment
    _overlay_coefficient_detail = deps.overlay_coefficient_detail
    _pick_market_reference = deps.pick_market_reference
    _overlay_model_leaderboard_row = deps.overlay_model_leaderboard_row
    _overlay_model_label = deps.overlay_model_label
    _driver_source_comment = deps.driver_source_comment
    _driver_source_note = deps.driver_source_note
    bridge_separator_rows: list[int] = []

    bridge_title_end_col = overlay_gpre_end_col if (is_gpre_profile and gpre_commercial_setup_rows) else max(2, 1 + len(overlay_display_quarters))
    bridge_end_col = max(2, 1 + len(overlay_display_quarters))

    def _bridge_label(label_in: str) -> str:
        label_txt = str(label_in or "").strip()
        if label_txt == "Total derivative P&L per gallon":
            return label_txt
        if is_gpre_profile and gpre_commercial_setup_rows and label_txt and "($m)" not in label_txt:
            return f"{label_txt} ($m)"
        return label_txt

    def _write_bridge_separator_row(separator_row_num: int) -> int:
        separator_fill = PatternFill(fill_type="solid", fgColor="EDF4FA")
        for cc in range(1, bridge_title_end_col + 1):
            ws.cell(row=separator_row_num, column=cc).fill = copy(separator_fill)
            ws.cell(row=separator_row_num, column=cc).border = Border()
        ws.row_dimensions[separator_row_num].height = 12.0
        bridge_separator_rows.append(separator_row_num)
        return separator_row_num + 1

    row_num = _write_section_bar(
        row_num,
        "Bridge to reported",
        end_col=bridge_title_end_col,
        primary=bool(is_gpre_profile and gpre_commercial_setup_rows),
        row_height=overlay_section_row_height if (is_gpre_profile and gpre_commercial_setup_rows) else None,
    )
    bridge_intro_end_col = 13 if (is_gpre_profile and gpre_commercial_setup_rows) else bridge_title_end_col
    row_num = _write_overlay_intro(
        row_num,
        "Approximate market crush shows simple weighted market/process conditions; GPRE crush proxy adds company-specific timing / hedge effects.",
        end_col=bridge_intro_end_col,
        spacer_after=1,
        row_height=overlay_intro_row_height if (is_gpre_profile and gpre_commercial_setup_rows) else None,
    )
    if is_gpre_profile and gpre_commercial_setup_rows:
        gpre_bridge_panel_rows["panel_title_row"] = row_num - 2
        gpre_bridge_panel_rows["panel_header_row"] = row_num - 1
        gpre_bridge_panel_rows["panel_subheader_row"] = row_num
    ws.cell(row=row_num, column=1, value="Quarter")
    ws.cell(row=row_num, column=1).font = bold_font
    ws.cell(row=row_num, column=1).fill = header_fill
    ws.cell(row=row_num, column=1).border = thin_border
    for cc in range(2, bridge_end_col + 1):
        ws.cell(row=row_num, column=cc).fill = header_fill
        ws.cell(row=row_num, column=cc).border = thin_border
        ws.cell(row=row_num, column=cc).alignment = Alignment(horizontal="center", vertical="center")
    for idx, qd in enumerate(overlay_display_quarters, start=2):
        qcell = ws.cell(row=row_num, column=idx, value=f"{qd.year}-Q{((qd.month - 1) // 3) + 1}")
        qcell.font = bold_font
        qcell.fill = header_fill
        qcell.border = thin_border
        qcell.alignment = Alignment(horizontal="center", vertical="center")
        if not (is_gpre_profile and gpre_commercial_setup_rows):
            ws.column_dimensions[get_column_letter(idx)].width = 14
    ws.row_dimensions[row_num].height = overlay_header_row_height if (is_gpre_profile and gpre_commercial_setup_rows) else 20
    row_num += 1

    bridge_key_map = {
        "reported_consolidated_crush_margin": "consolidated_ethanol_crush_margin",
        "underlying_crush_margin": "underlying_crush_margin",
        "base_business_adj_ebitda_ex_credits": "adjusted_ebitda_ex_45z_base_business",
    }
    bridge_label_overrides = {
        "approx_market_crush_proxy": "Approximate market crush",
        "gpre_crush_proxy": "GPRE crush proxy",
        "best_forward_lens_proxy": "Best forward lens",
        "45z": "45Z impact",
        "rin_sale": "RIN impact",
        "inventory_lcnrv": "Inventory NRV / lower-of-cost",
        "intercompany_nonethanol_net": "Non-ethanol operating activities",
        "impairment_assets_held_for_sale": "Impairment / held-for-sale",
        "other_bridge_items": "Other explicit bridge items",
        "base_business_adj_ebitda_ex_credits": "Base business Adj EBITDA ex-credits",
        "underlying_crush_margin": "Underlying crush margin",
        "reported_consolidated_crush_margin": "Reported consolidated crush margin",
        "total_derivative_pnl_per_gallon": "Total derivative P&L per gallon",
    }
    bridge_order = (
        "approx_market_crush_proxy",
        "gpre_crush_proxy",
        "best_forward_lens_proxy",
        "base_business_adj_ebitda_ex_credits",
        "underlying_crush_margin",
        "reported_consolidated_crush_margin",
        "total_derivative_pnl_per_gallon",
        "45z",
        "rin_sale",
        "inventory_lcnrv",
        "intercompany_nonethanol_net",
        "impairment_assets_held_for_sale",
        "other_bridge_items",
    )
    suppressed_bridge_keys = {"gap_vs_market_process_proxy", "hedge_realization_residual"}
    bridge_tpl_map = {str(getattr(tpl, "key", "") or ""): tpl for tpl in bridge_templates}
    market_tpl_by_key = {
        str(getattr(tpl, "key", "") or "").strip(): tpl
        for tpl in market_input_templates
        if str(getattr(tpl, "key", "") or "").strip()
    }
    bridge_value_cache: Dict[Tuple[str, date], Tuple[Optional[float], str]] = {}

    def _bridge_gallon_basis(qd_in: date) -> Tuple[Optional[float], str, str]:
        gallon_basis_order = (
            (
                ("ethanol_gallons_produced", "ethanol gallons produced"),
                ("ethanol_gallons_sold", "ethanol gallons sold"),
            )
            if is_gpre_profile and gpre_commercial_setup_rows
            else (
                ("ethanol_gallons_sold", "ethanol gallons sold"),
                ("ethanol_gallons_produced", "ethanol gallons produced"),
            )
        )
        for row_key, label_txt in gallon_basis_order:
            rec = row_map.get((row_key, qd_in))
            val_num = pd.to_numeric((rec or {}).get("Value"), errors="coerce")
            if pd.notna(val_num) and float(val_num) != 0.0:
                return float(val_num), label_txt, _driver_source_comment(rec)
        corn_consumed_num = pd.to_numeric((row_map.get(("corn_consumed", qd_in)) or {}).get("Value"), errors="coerce")
        ethanol_yield_num = pd.to_numeric((_overlay_coefficient_detail("ethanol_yield") or {}).get("value"), errors="coerce")
        if pd.notna(corn_consumed_num) and pd.notna(ethanol_yield_num):
            inferred_gallons = float(corn_consumed_num) * float(ethanol_yield_num)
            if inferred_gallons != 0.0:
                return (
                    inferred_gallons,
                    "estimated gallons from corn consumed and ethanol yield",
                    "Fallback gallon basis inferred from corn consumed and the platform ethanol-yield assumption.",
                )
        return None, "", ""

    def _bridge_value_to_per_gal(
        raw_value: Optional[float],
        qd_in: date,
        *,
        comment_txt: str,
    ) -> Tuple[Optional[float], str]:
        if raw_value is None:
            return None, ""
        gallon_basis, basis_label, basis_comment = _bridge_gallon_basis(qd_in)
        if gallon_basis is None or abs(float(gallon_basis)) < 1e-9:
            return None, ""
        parts = [
            str(comment_txt or "").strip(),
            f"Converted to $/gal using {basis_label}.",
            str(basis_comment or "").strip(),
        ]
        return float(raw_value) / float(gallon_basis), " ".join(part for part in parts if part)

    def _core_bridge_value(bkey_in: str, qd_in: date) -> Tuple[Optional[float], str]:
        mapped = bridge_key_map.get(bkey_in, "")
        if mapped:
            rec = row_map.get((mapped, qd_in))
            val_num = pd.to_numeric((rec or {}).get("Value"), errors="coerce")
            if pd.notna(val_num):
                comment_txt = _driver_source_comment(rec)
                if bkey_in == "base_business_adj_ebitda_ex_credits":
                    comment_txt = " ".join(
                        part
                        for part in (
                            "Base business = Adjusted EBITDA ex-45Z contribution. Underlying crush margin = reported consolidated crush margin ex-45Z COGS/crush benefit. These are different bridges.",
                            comment_txt,
                        )
                        if str(part or "").strip()
                    )
                return float(val_num), comment_txt
            return None, ""
        consolidated_rec = row_map.get(("consolidated_ethanol_crush_margin", qd_in))
        consolidated_num = pd.to_numeric((consolidated_rec or {}).get("Value"), errors="coerce")
        ex45z_rec = row_map.get(("crush_margin_ex_45z", qd_in))
        ex45z_num = pd.to_numeric((ex45z_rec or {}).get("Value"), errors="coerce")
        exrin_rec = row_map.get(("crush_margin_ex_rin", qd_in))
        exrin_num = pd.to_numeric((exrin_rec or {}).get("Value"), errors="coerce")
        bundle = bridge_bundle_map.get(qd_in) or {}
        comps = dict(bundle.get("components") or {})
        val = None
        comment_txt = ""
        if bkey_in == "45z" and pd.notna(consolidated_num) and pd.notna(ex45z_num):
            val = float(consolidated_num) - float(ex45z_num)
            comment_txt = _driver_source_note(
                (consolidated_rec or {}).get("source_doc"),
                "Derived as reported consolidated crush margin less crush margin ex-45Z.",
                _driver_source_comment(ex45z_rec),
            )
        elif bkey_in == "rin_sale" and pd.notna(consolidated_num) and pd.notna(exrin_num):
            val = float(consolidated_num) - float(exrin_num)
            comment_txt = _driver_source_note(
                (consolidated_rec or {}).get("source_doc"),
                "Derived as reported consolidated crush margin less crush margin ex-RIN.",
                _driver_source_comment(exrin_rec),
            )
        elif bkey_in == "other_bridge_items":
            known_keys = {"consolidated", "underlying", "ex_45z", "ex_rin", "45z", "rin_sale", "inventory_lcnrv", "intercompany_nonethanol_net", "impairment_assets_held_for_sale"}
            vals = [float(v) for k, v in comps.items() if k not in known_keys]
            if vals:
                val = float(sum(vals))
        elif bkey_in == "inventory_lcnrv" and is_gpre_profile and qd_in >= date(2026, 3, 31):
            # The Q1 2026 release table can place the prior-year inventory NRV value
            # next to the current-quarter 45Z bridge. Keep Q1 2026 blank unless a
            # same-quarter inventory NRV adjustment is separately sourced.
            val = None
        else:
            raw_val = comps.get(bkey_in)
            if raw_val is not None:
                val = float(raw_val)
        if val is not None and abs(float(val)) >= 200.0:
            val = float(val) / 1000.0
        if val is not None and not comment_txt:
            comment_txt = _driver_source_note(bundle.get("source_doc"), bundle.get("text"))
        return (float(val), comment_txt) if val is not None else (None, "")

    def _approx_market_proxy_value(qd_in: date) -> Tuple[Optional[float], str]:
        if is_gpre_profile and gpre_commercial_setup_rows:
            basis_rec = dict(gpre_basis_quarter_map.get(qd_in) or {})
            proxy_per_gal = pd.to_numeric(
                basis_rec.get(
                    "official_simple_proxy_usd_per_gal",
                    basis_rec.get("simple_market_proxy_usd_per_gal", basis_rec.get("official_proxy_usd_per_gal")),
                ),
                errors="coerce",
            )
            if pd.notna(proxy_per_gal):
                gallon_basis, _, _ = _bridge_gallon_basis(qd_in)
                if gallon_basis is None or abs(float(gallon_basis)) < 1e-9:
                    return None, ""
                comment_parts = [
                    "Approximate market crush uses the official simple GPRE market/process proxy.",
                    "Formula: weighted ethanol benchmark less delivered corn (CBOT corn + official weighted corn basis) and fixed gas burden, converted into the same bridge basis.",
                ]
                return float(proxy_per_gal) * float(gallon_basis), " ".join(part for part in comment_parts if part)
        corn_consumed_num = pd.to_numeric((row_map.get(("corn_consumed", qd_in)) or {}).get("Value"), errors="coerce")
        if pd.isna(corn_consumed_num):
            return None, ""
        needed_coeffs = {
            key: pd.to_numeric((_overlay_coefficient_detail(key) or {}).get("value"), errors="coerce")
            for key in ("ethanol_yield", "natural_gas_usage")
        }
        if any(pd.isna(val) for val in needed_coeffs.values()):
            return None, ""
        needed_inputs: Dict[str, float] = {}
        for input_key in ("corn_price", "ethanol_price", "natural_gas_price"):
            tpl = market_tpl_by_key.get(input_key)
            market_ref = _pick_market_reference(tpl, target_quarter=qd_in) if tpl is not None else None
            val_num = pd.to_numeric((market_ref or {}).get("_converted_value"), errors="coerce")
            if pd.isna(val_num):
                return None, ""
            needed_inputs[input_key] = float(val_num)
        process_per_bushel = (
            float(needed_coeffs["ethanol_yield"]) * float(needed_inputs["ethanol_price"])
            - float(needed_inputs["corn_price"])
            - (float(needed_coeffs["natural_gas_usage"]) / 1_000_000.0) * float(needed_coeffs["ethanol_yield"]) * float(needed_inputs["natural_gas_price"])
        )
        proxy_val = process_per_bushel * float(corn_consumed_num)
        return (
            float(proxy_val),
            "Approximate market crush from quarter-average weighted ethanol benchmark, delivered corn and gas inputs before hedge, policy and other bridge effects.",
        )

    def _gpre_proxy_role_value(
        qd_in: date,
        *,
        model_key: str,
        role_label: str,
    ) -> Tuple[Optional[float], str]:
        if not (is_gpre_profile and gpre_commercial_setup_rows):
            return None, ""
        basis_rec = dict(gpre_basis_quarter_map.get(qd_in) or {})
        resolved_model_key = str(model_key or current_overlay_model_key).strip()
        pred_col = (
            "gpre_proxy_official_usd_per_gal"
            if not resolved_model_key or resolved_model_key == current_overlay_model_key
            else str(overlay_model_key_to_pred_col.get(resolved_model_key) or "").strip()
        )
        proxy_per_gal = pd.to_numeric(basis_rec.get(pred_col), errors="coerce")
        if pd.isna(proxy_per_gal) and pred_col != "gpre_proxy_official_usd_per_gal":
            proxy_per_gal = pd.to_numeric(
                basis_rec.get(
                    "gpre_proxy_official_usd_per_gal",
                    basis_rec.get("bridge_official_proxy_usd_per_gal"),
                ),
                errors="coerce",
            )
        if pd.isna(proxy_per_gal):
            return None, ""
        gallon_basis, _, _ = _bridge_gallon_basis(qd_in)
        if gallon_basis is None or abs(float(gallon_basis)) < 1e-9:
            return None, ""
        role_row = _overlay_model_leaderboard_row(resolved_model_key)
        comment_parts = [
            f"{role_label} uses the {_overlay_model_label(resolved_model_key)} fitted proxy row.",
        ]
        family_txt = str(role_row.get("family_label") or role_row.get("family") or "").strip()
        timing_txt = str(role_row.get("timing_rule") or "").strip()
        clean_mae = pd.to_numeric(role_row.get("clean_mae"), errors="coerce")
        hybrid_score = pd.to_numeric(role_row.get("hybrid_score"), errors="coerce")
        forward_rating = str(role_row.get("forward_usability_rating") or "").strip()
        if family_txt or timing_txt:
            comment_parts.append(
                "Chosen model: "
                + " | ".join(part for part in (family_txt, timing_txt) if part)
                + "."
            )
        if pd.notna(clean_mae) or pd.notna(hybrid_score) or forward_rating:
            metric_bits = []
            if pd.notna(clean_mae):
                metric_bits.append(f"Clean MAE {float(clean_mae):.4f} $/gal")
            if pd.notna(hybrid_score):
                metric_bits.append(f"Hybrid {float(hybrid_score):.4f}")
            if forward_rating:
                metric_bits.append(f"Forward {forward_rating}")
            comment_parts.append("; ".join(metric_bits) + ".")
        return float(proxy_per_gal) * float(gallon_basis), " ".join(part for part in comment_parts if part)

    def _gpre_crush_proxy_value(qd_in: date) -> Tuple[Optional[float], str]:
        return _gpre_proxy_role_value(
            qd_in,
            model_key=current_overlay_model_key,
            role_label="GPRE crush proxy",
        )

    def _best_forward_lens_proxy_value(qd_in: date) -> Tuple[Optional[float], str]:
        return _gpre_proxy_role_value(
            qd_in,
            model_key=best_forward_overlay_model_key or current_overlay_model_key,
            role_label="Best forward lens",
        )

    def _bridge_value_and_comment(bkey_in: str, qd_in: date) -> Tuple[Optional[float], str]:
        cache_key = (bkey_in, qd_in)
        if cache_key in bridge_value_cache:
            return bridge_value_cache[cache_key]
        if bkey_in == "approx_market_crush_proxy":
            result = _approx_market_proxy_value(qd_in)
        elif bkey_in == "gpre_crush_proxy":
            result = _gpre_crush_proxy_value(qd_in)
        elif bkey_in == "best_forward_lens_proxy":
            result = _best_forward_lens_proxy_value(qd_in)
        elif bkey_in == "gap_vs_market_process_proxy":
            reported_val, _ = _bridge_value_and_comment("reported_consolidated_crush_margin", qd_in)
            proxy_val, _ = _bridge_value_and_comment("approx_market_crush_proxy", qd_in)
            result = (
                (float(reported_val) - float(proxy_val), "Reported consolidated crush margin less approximate market/process proxy.")
                if reported_val is not None and proxy_val is not None
                else (None, "")
            )
        elif bkey_in == "hedge_realization_residual":
            underlying_val, _ = _bridge_value_and_comment("underlying_crush_margin", qd_in)
            proxy_val, _ = _bridge_value_and_comment("approx_market_crush_proxy", qd_in)
            if underlying_val is not None and proxy_val is not None:
                result = (
                    float(underlying_val) - float(proxy_val),
                    "Residual between approximate market/process proxy and underlying crush margin.",
                )
            else:
                reported_val, _ = _bridge_value_and_comment("reported_consolidated_crush_margin", qd_in)
                explicit_vals = []
                for explicit_key in ("45z", "rin_sale", "inventory_lcnrv", "intercompany_nonethanol_net", "impairment_assets_held_for_sale", "other_bridge_items"):
                    explicit_val, _ = _core_bridge_value(explicit_key, qd_in)
                    if explicit_val is not None:
                        explicit_vals.append(float(explicit_val))
                result = (
                    (float(reported_val) - float(proxy_val) - float(sum(explicit_vals)), "Residual after explicit bridge items.")
                    if reported_val is not None and proxy_val is not None
                    else (None, "")
                )
        elif bkey_in == "total_derivative_pnl_per_gallon":
            der_rec = _derivative_bridge_record(derivative_bridge_by_quarter, qd_in)
            pnl_m = _derivative_usd_to_millions(der_rec.get("derivative_gain_loss_pnl_total_usd"))
            gallon_basis, basis_label, basis_comment = _bridge_gallon_basis(qd_in)
            if pnl_m is not None and gallon_basis is not None and abs(float(gallon_basis)) >= 1e-9:
                result = (
                    float(pnl_m) / float(gallon_basis),
                    " ".join(
                        part
                        for part in (
                            "Reported-margin-equivalent diagnostic; not pure spot crush margin.",
                            f"Converted to $/gal using {basis_label}.",
                            str(basis_comment or "").strip(),
                        )
                        if str(part or "").strip()
                    ),
                )
            else:
                result = (None, "")
        else:
            result = _core_bridge_value(bkey_in, qd_in)
        if is_gpre_profile and gpre_commercial_setup_rows:
            val_num = pd.to_numeric(result[0], errors="coerce")
            result = (float(val_num), result[1]) if pd.notna(val_num) else (None, result[1])
        else:
            result = _bridge_value_to_per_gal(result[0], qd_in, comment_txt=result[1])
        bridge_value_cache[cache_key] = result
        return result

    if is_gpre_profile and gpre_commercial_setup_rows:
        for qd in overlay_display_quarters:
            _, basis_label, _ = _bridge_gallon_basis(qd)
            if basis_label:
                gpre_denominator_policy_by_quarter[qd] = basis_label

    ordered_bridge_keys = [key for key in bridge_order if key in bridge_label_overrides]
    ordered_bridge_keys.extend(
        key for key in bridge_tpl_map.keys()
        if key and key not in ordered_bridge_keys and key not in suppressed_bridge_keys
    )

    for bkey in ordered_bridge_keys:
        label = _bridge_label(str(bridge_label_overrides.get(bkey) or getattr(bridge_tpl_map.get(bkey), "label", "") or bkey))
        if is_gpre_profile and gpre_commercial_setup_rows:
            gpre_bridge_panel_rows[bkey] = row_num
            if bkey == "base_business_adj_ebitda_ex_credits":
                gpre_bridge_panel_rows["proxy_implied_gallons"] = row_num
            elif bkey == "underlying_crush_margin":
                gpre_bridge_panel_rows["proxy_implied_volume_basis"] = row_num
        ws.cell(row=row_num, column=1, value=label)
        if is_gpre_profile and gpre_commercial_setup_rows and bkey == "base_business_adj_ebitda_ex_credits":
            _add_comment(
                f"A{row_num}",
                "Company-level Adj EBITDA bridge, ex-45Z.",
            )
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row_num, column=1).font = body_font
        for idx in range(2, bridge_end_col + 1):
            ws.cell(row=row_num, column=idx).border = thin_border
            ws.cell(row=row_num, column=idx).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_num, column=idx).font = body_font
        for idx, qd in enumerate(overlay_display_quarters, start=2):
            cell = ws.cell(row=row_num, column=idx)
            val, comment_txt = _bridge_value_and_comment(bkey, qd)
            if val is not None:
                cell.value = float(val)
                if is_gpre_profile and gpre_commercial_setup_rows and bkey == "total_derivative_pnl_per_gallon":
                    cell.number_format = "$0.000;($0.000);-"
                else:
                    cell.number_format = "0.0;-0.0" if (is_gpre_profile and gpre_commercial_setup_rows) else "0.000;-0.000"
                if is_gpre_profile and gpre_commercial_setup_rows and bkey == "reported_consolidated_crush_margin":
                    gpre_reported_margin_by_quarter[qd] = float(val)
        ws.row_dimensions[row_num].height = overlay_support_row_height if (is_gpre_profile and gpre_commercial_setup_rows) else 18
        row_num += 1
        if is_gpre_profile and gpre_commercial_setup_rows and bkey in {"best_forward_lens_proxy", "reported_consolidated_crush_margin"}:
            row_num = _write_bridge_separator_row(row_num)
    if is_gpre_profile and gpre_commercial_setup_rows and derivative_bridge_by_quarter:
        row_num = _write_section_bar(
            row_num,
            "Derivative / hedge memo",
            end_col=bridge_title_end_col,
            primary=False,
            row_height=overlay_section_row_height,
        )
        ws.cell(row=row_num, column=1, value="Quarter")
        ws.cell(row=row_num, column=1).font = bold_font
        ws.cell(row=row_num, column=1).fill = header_fill
        ws.cell(row=row_num, column=1).border = thin_border
        for idx, qd in enumerate(overlay_display_quarters, start=2):
            cell = ws.cell(row=row_num, column=idx, value=f"{qd.year}-Q{((qd.month - 1) // 3) + 1}")
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_num].height = overlay_header_row_height
        row_num += 1
        derivative_memo_rows = (
            ("Total derivative P&L", "derivative_gain_loss_pnl_total_usd", "Already embedded in revenue/COGS; memo only."),
            ("Derivative P&L in revenue", "derivative_gain_loss_revenue_usd", ""),
            ("Derivative P&L in COGS", "derivative_gain_loss_cogs_usd", ""),
            ("Cash-flow hedge reclass to P&L", "cash_flow_hedge_reclass_total_usd", "Pre-tax cash-flow hedge reclassification into revenue/COGS."),
            ("Net derivative asset/liability", "derivative_net_asset_liability_usd", "Balance-sheet derivative exposure at period end."),
            ("Derivative OCI movement", "derivative_oci_current_period_usd", ""),
            ("Derivative AOCI", "derivative_aoci_ending_balance_usd", "Accumulated cash-flow hedge OCI balance in equity."),
        )
        memo_fill = copy(zebra_fill_light)
        for label_txt, field_name, comment_txt in derivative_memo_rows:
            ws.cell(row=row_num, column=1, value=f"{label_txt} ($m)")
            ws.cell(row=row_num, column=1).font = body_font
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row=row_num, column=1).border = thin_border
            ws.cell(row=row_num, column=1).fill = copy(memo_fill)
            for idx, qd in enumerate(overlay_display_quarters, start=2):
                cell = ws.cell(row=row_num, column=idx)
                der_rec = _derivative_bridge_record(derivative_bridge_by_quarter, qd)
                val_m = _derivative_usd_to_millions(der_rec.get(field_name))
                if val_m is not None:
                    cell.value = float(val_m)
                    cell.number_format = "0.0;-0.0"
                cell.font = body_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                cell.fill = copy(memo_fill)
            ws.row_dimensions[row_num].height = overlay_support_row_height
            row_num += 1
            if label_txt == "Cash-flow hedge reclass to P&L":
                row_num = _write_bridge_separator_row(row_num)

    return GpreEconomicsOverlayBridgeResult(
        row_idx=row_num,
        bridge_separator_rows=bridge_separator_rows,
        gpre_bridge_panel_rows=gpre_bridge_panel_rows,
        gpre_reported_margin_by_quarter=gpre_reported_margin_by_quarter,
        gpre_denominator_policy_by_quarter=gpre_denominator_policy_by_quarter,
    )
