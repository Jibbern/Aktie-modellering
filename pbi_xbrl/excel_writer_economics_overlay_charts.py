"""Economics_Overlay chart writer helpers."""
from __future__ import annotations

import time
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any, Callable, Dict, List, Mapping, Optional, Sequence, Set, Tuple

import pandas as pd
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties, LineProperties
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel


@dataclass(frozen=True)
class EconomicsOverlayChartWriterDeps:
    ws: Any
    row_idx: int
    is_gpre_profile: bool
    gpre_commercial_setup_rows: Sequence[Mapping[str, Any]]
    simple_crush_history_rows: Sequence[Mapping[str, Any]]
    proxy_comp_end_row: int
    official_proxy_comp_row: int
    fitted_proxy_comp_row: int
    best_forward_proxy_comp_row: int
    next_quarter_thesis_snapshot: Mapping[str, Any]
    sandbox_process_margin_refs: Mapping[str, Any]
    thesis_ref: Mapping[str, str]
    prior_market_display_quarter: Any
    quarter_open_display_quarter: Any
    current_market_display_quarter: Any
    next_thesis_quarter_end: Any
    quarterly_df: Any
    overlay_model_key_to_pred_col: Mapping[str, str]
    current_overlay_model_key: str
    best_forward_overlay_model_key: str
    title_fill: Any
    title_font: Any
    thin_border: Any
    align_center: Any
    add_comment: Callable[[str, Any], None]
    gpre_preview_frame_value: Callable[..., Any]
    gpre_model_preview_frame_value: Callable[..., Any]
    historical_proxy_value: Callable[..., Any]
    apply_chart_text_categories: Callable[..., None]
    record_writer_substage: Callable[[str, float], None]
    chart_width: float = 34.0
    chart_height: float = 16.0
    chart_row_span: int = 24
    max_chart_points: int = 15


@dataclass(frozen=True)
class EconomicsOverlayChartWriterResult:
    row_idx: int


def _quarter_bounds_from_end_date(q_end_in: Any) -> Tuple[Optional[date], Optional[date]]:
    if not isinstance(q_end_in, date):
        return None, None
    try:
        q_period = pd.Timestamp(q_end_in).to_period("Q")
        q_start = q_period.start_time.normalize().date()
        q_end = q_period.end_time.normalize().date()
        return q_start, q_end
    except Exception:
        month_num = ((int(q_end_in.month) - 1) // 3) * 3 + 1
        q_start = date(int(q_end_in.year), month_num, 1)
        return q_start, q_end_in


def _build_visible_quarter_label_points(
    visible_start: Optional[date],
    visible_end: Optional[date],
) -> List[Dict[str, Any]]:
    if not isinstance(visible_start, date) or not isinstance(visible_end, date) or visible_end < visible_start:
        return []
    current_start, _ = _quarter_bounds_from_end_date(visible_start)
    if not isinstance(current_start, date):
        return []
    out: List[Dict[str, Any]] = []
    while isinstance(current_start, date) and current_start <= visible_end:
        try:
            current_period = pd.Timestamp(current_start).to_period("Q")
            quarter_start = current_period.start_time.normalize().date()
            quarter_end = current_period.end_time.normalize().date()
        except Exception:
            month_num = ((int(current_start.month) - 1) // 3) * 3 + 1
            quarter_start = date(int(current_start.year), month_num, 1)
            if month_num == 10:
                quarter_end = date(int(current_start.year), 12, 31)
            else:
                quarter_end = date(int(current_start.year), month_num + 3, 1) - timedelta(days=1)
        clip_start = max(quarter_start, visible_start)
        clip_end = min(quarter_end, visible_end)
        if clip_end >= clip_start:
            midpoint = clip_start + timedelta(days=max(((clip_end - clip_start).days // 2), 0))
            q_num = ((int(quarter_end.month) - 1) // 3) + 1
            out.append(
                {
                    "quarter_start": quarter_start,
                    "quarter_end": quarter_end,
                    "clip_start": clip_start,
                    "clip_end": clip_end,
                    "midpoint": midpoint,
                    "label": f"{int(quarter_end.year)}-Q{q_num}",
                }
            )
        current_start = quarter_end + timedelta(days=1)
    return out


def write_economics_overlay_charts(
    deps: EconomicsOverlayChartWriterDeps,
) -> EconomicsOverlayChartWriterResult:
    ws = deps.ws
    row_idx = int(deps.row_idx)
    is_gpre_profile = deps.is_gpre_profile
    gpre_commercial_setup_rows = deps.gpre_commercial_setup_rows
    simple_crush_history_rows = deps.simple_crush_history_rows
    proxy_comp_end_row = int(deps.proxy_comp_end_row or 0)
    official_proxy_comp_row = int(deps.official_proxy_comp_row or 0)
    fitted_proxy_comp_row = int(deps.fitted_proxy_comp_row or 0)
    best_forward_proxy_comp_row = int(deps.best_forward_proxy_comp_row or 0)
    next_quarter_thesis_snapshot = deps.next_quarter_thesis_snapshot
    sandbox_process_margin_refs = deps.sandbox_process_margin_refs
    thesis_ref = deps.thesis_ref
    prior_market_display_quarter = deps.prior_market_display_quarter
    quarter_open_display_quarter = deps.quarter_open_display_quarter
    current_market_display_quarter = deps.current_market_display_quarter
    next_thesis_quarter_end = deps.next_thesis_quarter_end
    quarterly_df = deps.quarterly_df
    overlay_model_key_to_pred_col = deps.overlay_model_key_to_pred_col
    current_overlay_model_key = deps.current_overlay_model_key
    best_forward_overlay_model_key = deps.best_forward_overlay_model_key
    title_fill = deps.title_fill
    title_font = deps.title_font
    thin_border = deps.thin_border
    align_center = deps.align_center
    _add_comment = deps.add_comment
    _gpre_preview_frame_value = deps.gpre_preview_frame_value
    _gpre_model_preview_frame_value = deps.gpre_model_preview_frame_value
    _historical_proxy_value = deps.historical_proxy_value
    _apply_chart_text_categories = deps.apply_chart_text_categories
    _record_writer_substage = deps.record_writer_substage
    overlay_charts_started = time.perf_counter()
    # Keep overlay charts on one visual grid and cap quarterly/coprod windows so
    # new quarters roll in without turning the delivered workbook into a horizontal dump.
    overlay_chart_width = float(deps.chart_width)
    overlay_chart_height = float(deps.chart_height)
    overlay_chart_row_span = int(deps.chart_row_span)
    overlay_quarter_chart_max_points = int(deps.max_chart_points)
    if is_gpre_profile and gpre_commercial_setup_rows and len(simple_crush_history_rows) >= 2:
        chart_title_row = max(proxy_comp_end_row + 2, row_idx)
        while bool(ws.row_dimensions[chart_title_row].hidden):
            chart_title_row += 1
        chart_anchor_row = chart_title_row + 1
        chart_end_row = chart_anchor_row + overlay_chart_row_span
        chart_start_col = 2  # B
        chart_end_col = 21  # U
        history_col_date = 37  # AK
        history_col_value = 38  # AL
        thesis_col_date = 39  # AM
        thesis_col_value = 40  # AN
        prior_preview_col_date = 41  # AO
        prior_preview_col_value = 42  # AP
        quarter_open_preview_col_date = 43  # AQ
        quarter_open_preview_col_value = 44  # AR
        boundary_col_date = 51  # AY
        boundary_col_value = 52  # AZ
        label_col_date = 53  # BA
        label_col_value = 54  # BB
        legend_swatch_col_date = 55  # BC
        legend_swatch_col_value = 56  # BD
        legend_label_col_date = 57  # BE
        legend_label_col_value = 58  # BF
        ws.column_dimensions[get_column_letter(history_col_date)].hidden = True
        ws.column_dimensions[get_column_letter(history_col_value)].hidden = True
        ws.column_dimensions[get_column_letter(thesis_col_date)].hidden = True
        ws.column_dimensions[get_column_letter(thesis_col_value)].hidden = True
        ws.column_dimensions[get_column_letter(prior_preview_col_date)].hidden = True
        ws.column_dimensions[get_column_letter(prior_preview_col_value)].hidden = True
        ws.column_dimensions[get_column_letter(quarter_open_preview_col_date)].hidden = True
        ws.column_dimensions[get_column_letter(quarter_open_preview_col_value)].hidden = True
        ws.column_dimensions[get_column_letter(boundary_col_date)].hidden = True
        ws.column_dimensions[get_column_letter(boundary_col_value)].hidden = True
        ws.column_dimensions[get_column_letter(label_col_date)].hidden = True
        ws.column_dimensions[get_column_letter(label_col_value)].hidden = True
        ws.column_dimensions[get_column_letter(legend_swatch_col_date)].hidden = True
        ws.column_dimensions[get_column_letter(legend_swatch_col_value)].hidden = True
        ws.column_dimensions[get_column_letter(legend_label_col_date)].hidden = True
        ws.column_dimensions[get_column_letter(legend_label_col_value)].hidden = True
        helper_start_row = chart_title_row
        ws.cell(row=helper_start_row, column=history_col_date, value="Week ending")
        ws.cell(row=helper_start_row, column=history_col_value, value="Approximate market crush ($/gal)")
        ws.cell(row=helper_start_row, column=thesis_col_date, value="Thesis week ending")
        ws.cell(row=helper_start_row, column=thesis_col_value, value="Next quarter outlook ($/gal)")
        ws.cell(row=helper_start_row, column=prior_preview_col_date, value="Prior quarter week ending")
        ws.cell(row=helper_start_row, column=prior_preview_col_value, value="Prior quarter ($/gal)")
        ws.cell(row=helper_start_row, column=quarter_open_preview_col_date, value="Quarter-open week ending")
        ws.cell(row=helper_start_row, column=quarter_open_preview_col_value, value="Qtr-open avg")
        ws.cell(row=helper_start_row, column=boundary_col_date, value="Quarter boundary")
        ws.cell(row=helper_start_row, column=boundary_col_value, value="Quarter boundary value")
        ws.cell(row=helper_start_row, column=label_col_date, value="Quarter label date")
        ws.cell(row=helper_start_row, column=label_col_value, value="Quarter label value")
        ws.cell(row=helper_start_row, column=legend_swatch_col_date, value="Legend swatch date")
        ws.cell(row=helper_start_row, column=legend_swatch_col_value, value="Legend swatch value")
        ws.cell(row=helper_start_row, column=legend_label_col_date, value="Legend label date")
        ws.cell(row=helper_start_row, column=legend_label_col_value, value="Legend label value")
        helper_last_row = helper_start_row
        history_dates: List[date] = []
        history_values: List[float] = []
        for idx, rec in enumerate(simple_crush_history_rows, start=helper_start_row + 1):
            week_end = rec.get("week_end")
            crush_per_gal = pd.to_numeric(rec.get("simple_crush_per_gal"), errors="coerce")
            if not isinstance(week_end, date) or pd.isna(crush_per_gal):
                continue
            ws.cell(row=idx, column=history_col_date, value=week_end)
            ws.cell(row=idx, column=history_col_date).number_format = "yyyy-mm-dd"
            ws.cell(row=idx, column=history_col_value, value=float(crush_per_gal))
            ws.cell(row=idx, column=history_col_value).number_format = "#,##0.000"
            helper_last_row = idx
            history_dates.append(week_end)
            history_values.append(float(crush_per_gal))
        if helper_last_row > helper_start_row:
            next_q_start = next_quarter_thesis_snapshot.get("target_quarter_start") if isinstance(next_quarter_thesis_snapshot, dict) else None
            next_q_end = next_quarter_thesis_snapshot.get("target_quarter_end") if isinstance(next_quarter_thesis_snapshot, dict) else None
            thesis_value_ref = str(sandbox_process_margin_refs.get("next_quarter_thesis") or "")
            thesis_ready_refs = [thesis_ref.get("corn_price", ""), thesis_ref.get("natural_gas_price", ""), thesis_ref.get("ethanol_price", ""), thesis_value_ref]
            thesis_checks = ",".join(f"ISNUMBER({ref})" for ref in thesis_ready_refs if ref)
            thesis_start_row = helper_start_row + 1
            thesis_mid_row = helper_start_row + 2
            thesis_end_row = helper_start_row + 3
            thesis_mid_date = None
            if isinstance(next_q_start, date) and isinstance(next_q_end, date):
                thesis_mid_date = next_q_start + timedelta(days=max(((next_q_end - next_q_start).days // 2), 0))
            thesis_date_rows = (
                (thesis_start_row, next_q_start),
                (thesis_mid_row, thesis_mid_date),
                (thesis_end_row, next_q_end),
            )
            for target_row, target_date in thesis_date_rows:
                if isinstance(target_date, date):
                    ws.cell(row=target_row, column=thesis_col_date, value=target_date)
                    ws.cell(row=target_row, column=thesis_col_date).number_format = "yyyy-mm-dd"
            thesis_formula = f'=IF(AND({thesis_checks}),{thesis_value_ref},NA())' if thesis_checks else '=NA()'
            for target_row in (thesis_start_row, thesis_mid_row, thesis_end_row):
                ws.cell(row=target_row, column=thesis_col_value, value=thesis_formula)
                ws.cell(row=target_row, column=thesis_col_value).number_format = "$0.000"

            prior_preview_start = None
            prior_preview_end = None
            quarter_open_preview_start = None
            quarter_open_preview_end = None
            if isinstance(prior_market_display_quarter, date):
                prior_preview_start, prior_preview_end = _quarter_bounds_from_end_date(prior_market_display_quarter)
            if isinstance(quarter_open_display_quarter, date):
                quarter_open_preview_start, quarter_open_preview_end = _quarter_bounds_from_end_date(quarter_open_display_quarter)
            prior_preview_mid = (
                prior_preview_start + timedelta(days=max(((prior_preview_end - prior_preview_start).days // 2), 0))
                if isinstance(prior_preview_start, date) and isinstance(prior_preview_end, date)
                else None
            )
            quarter_open_preview_mid = (
                quarter_open_preview_start + timedelta(days=max(((quarter_open_preview_end - quarter_open_preview_start).days // 2), 0))
                if isinstance(quarter_open_preview_start, date) and isinstance(quarter_open_preview_end, date)
                else None
            )
            prior_preview_formula = (
                f'=IF(ISNUMBER(B{official_proxy_comp_row}),B{official_proxy_comp_row},NA())'
                if official_proxy_comp_row > 0
                else '=NA()'
            )
            quarter_open_preview_formula = (
                f'=IF(ISNUMBER(D{official_proxy_comp_row}),D{official_proxy_comp_row},NA())'
                if official_proxy_comp_row > 0
                else '=NA()'
            )
            prior_preview_rows = (
                (helper_start_row + 1, prior_preview_start),
                (helper_start_row + 2, prior_preview_mid),
                (helper_start_row + 3, prior_preview_end),
            )
            quarter_open_preview_rows = (
                (helper_start_row + 1, quarter_open_preview_start),
                (helper_start_row + 2, quarter_open_preview_mid),
                (helper_start_row + 3, quarter_open_preview_end),
            )
            for target_row, target_date in prior_preview_rows:
                if isinstance(target_date, date):
                    ws.cell(row=target_row, column=prior_preview_col_date, value=target_date)
                    ws.cell(row=target_row, column=prior_preview_col_date).number_format = "yyyy-mm-dd"
                    ws.cell(row=target_row, column=prior_preview_col_value, value=prior_preview_formula)
                    ws.cell(row=target_row, column=prior_preview_col_value).number_format = "$0.000"
            for target_row, target_date in quarter_open_preview_rows:
                if isinstance(target_date, date):
                    ws.cell(row=target_row, column=quarter_open_preview_col_date, value=target_date)
                    ws.cell(row=target_row, column=quarter_open_preview_col_date).number_format = "yyyy-mm-dd"
                    ws.cell(row=target_row, column=quarter_open_preview_col_value, value=quarter_open_preview_formula)
                    ws.cell(row=target_row, column=quarter_open_preview_col_value).number_format = "$0.000"

            y_values = list(history_values)
            quarter_open_preview_val = _gpre_preview_frame_value("official_frames", "quarter_open")
            if quarter_open_preview_val is not None:
                y_values.append(float(quarter_open_preview_val))
            next_thesis_preview_val = _gpre_preview_frame_value("official_frames", "next_quarter_thesis")
            if next_thesis_preview_val is not None:
                y_values.append(float(next_thesis_preview_val))
            if y_values:
                y_min = min(y_values)
                y_max = max(y_values)
            else:
                y_min, y_max = (-0.50, 0.50)
            plotted_dates: List[date] = [hist_dt for hist_dt in history_dates if isinstance(hist_dt, date)]
            if next_thesis_preview_val is not None:
                for thesis_dt in (next_q_start, thesis_mid_date, next_q_end):
                    if isinstance(thesis_dt, date):
                        plotted_dates.append(thesis_dt)
            visible_start = None
            visible_end = None
            if plotted_dates:
                earliest_plotted = min(plotted_dates)
                latest_plotted = max(plotted_dates)
                visible_start, _ = _quarter_bounds_from_end_date(earliest_plotted)
                _, visible_end = _quarter_bounds_from_end_date(latest_plotted)
            visible_quarter_labels = _build_visible_quarter_label_points(visible_start, visible_end)

            ws.merge_cells(start_row=chart_title_row, start_column=chart_start_col, end_row=chart_title_row, end_column=chart_end_col)
            title_cell = ws.cell(row=chart_title_row, column=2, value="Approximate market crush (weekly)")
            title_cell.fill = title_fill
            title_cell.font = title_font
            title_cell.alignment = align_center
            title_cell.border = thin_border
            _add_comment(
                f"B{chart_title_row}",
                "The solid line is spot weekly approximate market crush. Dashed reference lines show quarter-open and next-quarter average lenses.",
            )
            for cc in range(chart_start_col, chart_end_col + 1):
                ws.cell(row=chart_title_row, column=cc).fill = title_fill
                ws.cell(row=chart_title_row, column=cc).font = title_font
                ws.cell(row=chart_title_row, column=cc).alignment = align_center
                ws.cell(row=chart_title_row, column=cc).border = thin_border
            ws.row_dimensions[chart_title_row].height = 18.0
            chart = ScatterChart()
            chart.scatterStyle = "line"
            chart.title = None
            chart.height = overlay_chart_height
            chart.width = overlay_chart_width
            chart.x_axis.title = None
            chart.x_axis.axPos = "b"
            chart.x_axis.delete = False
            chart.x_axis.number_format = ";;;"
            chart.x_axis.tickLblPos = "low"
            chart.x_axis.majorUnit = 120
            chart.x_axis.crosses = "min"
            chart.x_axis.majorTickMark = "none"
            chart.x_axis.minorTickMark = "none"
            try:
                chart.x_axis.majorGridlines = None
            except Exception:
                pass
            chart.y_axis.title = None
            chart.y_axis.axPos = "l"
            chart.y_axis.delete = False
            chart.y_axis.number_format = '$0.00'
            chart.y_axis.tickLblPos = "nextTo"
            chart.y_axis.crosses = "min"
            chart.y_axis.majorTickMark = "out"
            chart.y_axis.minorTickMark = "none"
            chart.legend = None
            # The chart source lives in hidden helper columns on the same sheet.
            # Excel skips hidden source cells unless the chart explicitly opts in.
            chart.visible_cells_only = False
            data_ref = Reference(ws, min_col=history_col_value, min_row=helper_start_row + 1, max_row=helper_last_row)
            x_ref = Reference(ws, min_col=history_col_date, min_row=helper_start_row + 1, max_row=helper_last_row)
            series = Series(data_ref, xvalues=x_ref, title="Approximate market crush ($/gal)")
            chart.series.append(series)
            try:
                chart.series[0].graphicalProperties.line.solidFill = "2F80ED"
                chart.series[0].graphicalProperties.line.width = 19050
                chart.series[0].marker.symbol = "none"
            except Exception:
                pass
            quarter_open_preview_data_ref = Reference(ws, min_col=quarter_open_preview_col_value, min_row=helper_start_row + 1, max_row=helper_start_row + 3)
            quarter_open_preview_x_ref = Reference(ws, min_col=quarter_open_preview_col_date, min_row=helper_start_row + 1, max_row=helper_start_row + 3)
            quarter_open_preview_series = Series(quarter_open_preview_data_ref, xvalues=quarter_open_preview_x_ref, title="Qtr-open avg")
            chart.series.append(quarter_open_preview_series)
            try:
                chart.series[1].graphicalProperties.line.solidFill = "E67E22"
                chart.series[1].graphicalProperties.line.width = 25400
                chart.series[1].graphicalProperties.line.dashStyle = "sysDash"
                chart.series[1].marker.symbol = "none"
            except Exception:
                pass
            thesis_data_ref = Reference(ws, min_col=thesis_col_value, min_row=thesis_start_row, max_row=thesis_end_row)
            thesis_x_ref = Reference(ws, min_col=thesis_col_date, min_row=thesis_start_row, max_row=thesis_end_row)
            thesis_series = Series(thesis_data_ref, xvalues=thesis_x_ref, title="Next qtr avg")
            chart.series.append(thesis_series)
            try:
                chart.series[2].graphicalProperties.line.solidFill = "27AE60"
                chart.series[2].graphicalProperties.line.width = 31750
                chart.series[2].graphicalProperties.line.dashStyle = "sysDash"
                chart.series[2].marker.symbol = "none"
            except Exception:
                pass
            y_span = max(float(y_max) - float(y_min), 0.10)
            y_pad = max(y_span * 0.05, 0.03)
            boundary_min = float(y_min) - (y_pad * 0.35)
            label_band = max(y_span * 0.12, 0.08)
            boundary_max = float(y_max) + (y_pad * 0.35)
            legend_band = max(y_span * 0.20, 0.12)
            label_y = boundary_min - (label_band * 0.45)
            chart_y_min = boundary_min - label_band
            chart_y_max = boundary_max + legend_band
            boundary_dates = [item.get("quarter_start") for item in visible_quarter_labels[1:]]
            boundary_dates = [bd for bd in boundary_dates if isinstance(bd, date)]
            boundary_start_row = max(helper_last_row, thesis_end_row) + 2
            boundary_row = boundary_start_row
            boundary_series_rows: List[Tuple[int, int]] = []
            for boundary_date in boundary_dates:
                line_start_row = boundary_row
                line_end_row = boundary_row + 1
                ws.cell(row=line_start_row, column=boundary_col_date, value=boundary_date)
                ws.cell(row=line_start_row, column=boundary_col_date).number_format = "yyyy-mm-dd"
                ws.cell(row=line_start_row, column=boundary_col_value, value=boundary_min)
                ws.cell(row=line_end_row, column=boundary_col_date, value=boundary_date)
                ws.cell(row=line_end_row, column=boundary_col_date).number_format = "yyyy-mm-dd"
                ws.cell(row=line_end_row, column=boundary_col_value, value=boundary_max)
                boundary_series_rows.append((line_start_row, line_end_row))
                boundary_row += 2
            for line_start_row, line_end_row in boundary_series_rows:
                boundary_data_ref = Reference(ws, min_col=boundary_col_value, min_row=line_start_row, max_row=line_end_row)
                boundary_x_ref = Reference(ws, min_col=boundary_col_date, min_row=line_start_row, max_row=line_end_row)
                # Leave helper guide series untitled. Passing title="" makes
                # openpyxl emit an empty <tx/> element, which Excel repairs by
                # dropping the whole drawing part.
                boundary_series = Series(boundary_data_ref, xvalues=boundary_x_ref)
                chart.series.append(boundary_series)
                try:
                    series_idx = len(chart.series) - 1
                    chart.series[series_idx].graphicalProperties.line.solidFill = "A0A0A0"
                    chart.series[series_idx].graphicalProperties.line.width = 8250
                    chart.series[series_idx].marker.symbol = "none"
                except Exception:
                    pass
            label_start_row = boundary_row + 1
            label_row = label_start_row
            for quarter_info in visible_quarter_labels:
                midpoint_date = quarter_info.get("midpoint")
                label_txt = str(quarter_info.get("label") or "").strip()
                if not isinstance(midpoint_date, date) or not label_txt:
                    continue
                ws.cell(row=label_row, column=label_col_date, value=midpoint_date)
                ws.cell(row=label_row, column=label_col_date).number_format = "yyyy-mm-dd"
                ws.cell(row=label_row, column=label_col_value, value=float(label_y))
                label_data_ref = Reference(ws, min_col=label_col_value, min_row=label_row, max_row=label_row)
                label_x_ref = Reference(ws, min_col=label_col_date, min_row=label_row, max_row=label_row)
                label_series = Series(label_data_ref, xvalues=label_x_ref, title=label_txt)
                chart.series.append(label_series)
                try:
                    series_idx = len(chart.series) - 1
                    chart.series[series_idx].graphicalProperties.line.solidFill = "FFFFFF"
                    chart.series[series_idx].graphicalProperties.line.width = 1
                    chart.series[series_idx].marker.symbol = "none"
                    chart.series[series_idx].dLbls = DataLabelList(showSerName=True, showVal=False, showCatName=False, dLblPos="ctr")
                except Exception:
                    pass
                label_row += 1
            # The built-in chart legend is all-or-nothing in several renderers.
            # Draw a tiny manual legend instead, so helper series can never
            # leak into the visible legend.
            legend_visible_start = visible_start or (min(history_dates) if history_dates else None)
            legend_visible_end = visible_end or (max(history_dates) if history_dates else None)
            if isinstance(legend_visible_start, date) and isinstance(legend_visible_end, date) and legend_visible_end > legend_visible_start:
                legend_range_days = max((legend_visible_end - legend_visible_start).days, 1)
                legend_swatch_start_date = legend_visible_start + timedelta(days=max(int(legend_range_days * 0.755), 1))
                legend_swatch_end_date = legend_visible_start + timedelta(days=max(int(legend_range_days * 0.805), 2))
                legend_label_date = legend_visible_start + timedelta(days=max(int(legend_range_days * 0.818), 3))
                legend_start_y = boundary_max + (legend_band * 0.82)
                legend_gap_y = legend_band * 0.30
                legend_specs = (
                    ("Qtr-open avg", "E67E22"),
                    ("Next qtr avg", "27AE60"),
                )
                legend_row = label_row + 1
                for legend_idx, (legend_label, legend_color) in enumerate(legend_specs):
                    legend_y = legend_start_y - (legend_idx * legend_gap_y)
                    swatch_start_row = legend_row
                    swatch_end_row = legend_row + 1
                    label_point_row = legend_row
                    ws.cell(row=swatch_start_row, column=legend_swatch_col_date, value=legend_swatch_start_date)
                    ws.cell(row=swatch_start_row, column=legend_swatch_col_date).number_format = "yyyy-mm-dd"
                    ws.cell(row=swatch_start_row, column=legend_swatch_col_value, value=float(legend_y))
                    ws.cell(row=swatch_end_row, column=legend_swatch_col_date, value=legend_swatch_end_date)
                    ws.cell(row=swatch_end_row, column=legend_swatch_col_date).number_format = "yyyy-mm-dd"
                    ws.cell(row=swatch_end_row, column=legend_swatch_col_value, value=float(legend_y))
                    ws.cell(row=label_point_row, column=legend_label_col_date, value=legend_label_date)
                    ws.cell(row=label_point_row, column=legend_label_col_date).number_format = "yyyy-mm-dd"
                    ws.cell(row=label_point_row, column=legend_label_col_value, value=float(legend_y))

                    swatch_data_ref = Reference(ws, min_col=legend_swatch_col_value, min_row=swatch_start_row, max_row=swatch_end_row)
                    swatch_x_ref = Reference(ws, min_col=legend_swatch_col_date, min_row=swatch_start_row, max_row=swatch_end_row)
                    swatch_series = Series(swatch_data_ref, xvalues=swatch_x_ref)
                    chart.series.append(swatch_series)
                    try:
                        series_idx = len(chart.series) - 1
                        chart.series[series_idx].graphicalProperties.line.solidFill = legend_color
                        chart.series[series_idx].graphicalProperties.line.width = 25400
                        chart.series[series_idx].graphicalProperties.line.dashStyle = "sysDash"
                        chart.series[series_idx].marker.symbol = "none"
                    except Exception:
                        pass

                    legend_label_data_ref = Reference(ws, min_col=legend_label_col_value, min_row=label_point_row, max_row=label_point_row)
                    legend_label_x_ref = Reference(ws, min_col=legend_label_col_date, min_row=label_point_row, max_row=label_point_row)
                    legend_label_series = Series(legend_label_data_ref, xvalues=legend_label_x_ref, title=legend_label)
                    chart.series.append(legend_label_series)
                    try:
                        series_idx = len(chart.series) - 1
                        chart.series[series_idx].graphicalProperties.line.solidFill = "FFFFFF"
                        chart.series[series_idx].graphicalProperties.line.width = 1
                        chart.series[series_idx].marker.symbol = "none"
                        chart.series[series_idx].dLbls = DataLabelList(showSerName=True, showVal=False, showCatName=False, dLblPos="r")
                    except Exception:
                        pass
                    legend_row += 2
            chart.legend = None
            ws.row_dimensions[chart_title_row].hidden = False
            if isinstance(visible_start, date):
                try:
                    chart.x_axis.scaling.min = float(to_excel(datetime.combine(visible_start, datetime.min.time())))
                except Exception:
                    pass
            if isinstance(visible_end, date):
                try:
                    chart.x_axis.scaling.max = float(to_excel(datetime.combine(visible_end + timedelta(days=1), datetime.min.time())))
                except Exception:
                    pass
            try:
                chart.y_axis.scaling.min = float(chart_y_min)
                chart.y_axis.scaling.max = float(chart_y_max)
            except Exception:
                pass
            chart.anchor = TwoCellAnchor(
                _from=AnchorMarker(col=1, row=chart_anchor_row - 1),
                to=AnchorMarker(col=chart_end_col, row=chart_end_row),
            )
            ws.add_chart(chart)

            quarterly_chart_rows_by_end: Dict[date, Dict[str, Any]] = {}
            historical_quarter_ends: Set[date] = set()
            if isinstance(quarterly_df, pd.DataFrame) and not quarterly_df.empty:
                quarter_series = pd.to_datetime(quarterly_df.get("quarter"), errors="coerce").dt.date
                official_series = (
                    pd.to_numeric(quarterly_df.get("official_simple_proxy_usd_per_gal"), errors="coerce")
                    if "official_simple_proxy_usd_per_gal" in quarterly_df.columns
                    else pd.Series(dtype=float)
                )
                fitted_series = (
                    pd.to_numeric(quarterly_df.get("gpre_proxy_official_usd_per_gal"), errors="coerce")
                    if "gpre_proxy_official_usd_per_gal" in quarterly_df.columns
                    else pd.Series(dtype=float)
                )
                best_forward_pred_col = str(
                    overlay_model_key_to_pred_col.get(best_forward_overlay_model_key or current_overlay_model_key) or ""
                ).strip()
                best_forward_series = (
                    pd.to_numeric(quarterly_df.get(best_forward_pred_col), errors="coerce")
                    if best_forward_pred_col and best_forward_pred_col in quarterly_df.columns
                    else fitted_series
                )
                realized_consolidated_series = (
                    pd.to_numeric(quarterly_df.get("reported_consolidated_crush_margin_usd_per_gal"), errors="coerce")
                    if "reported_consolidated_crush_margin_usd_per_gal" in quarterly_df.columns
                    else pd.Series(dtype=float)
                )
                realized_underlying_series = (
                    pd.to_numeric(quarterly_df.get("underlying_crush_margin_usd_per_gal"), errors="coerce")
                    if "underlying_crush_margin_usd_per_gal" in quarterly_df.columns
                    else pd.Series(dtype=float)
                )
                # The realized quarterly series intentionally changes lens at 2025-Q2:
                # reported consolidated before the cutover, underlying from the cutover onward.
                realized_cutover_quarter_end = date(2025, 6, 30)
                for quarter_end, official_num, fitted_num, best_forward_num, consolidated_num, underlying_num in zip(
                    quarter_series,
                    official_series,
                    fitted_series,
                    best_forward_series,
                    realized_consolidated_series,
                    realized_underlying_series,
                ):
                    if not isinstance(quarter_end, date):
                        continue
                    selected_realized_num = underlying_num if quarter_end >= realized_cutover_quarter_end else consolidated_num
                    if pd.isna(official_num) and pd.isna(fitted_num) and pd.isna(best_forward_num) and pd.isna(selected_realized_num):
                        continue
                    historical_quarter_ends.add(quarter_end)
                    quarterly_chart_rows_by_end[quarter_end] = {
                        "quarter_end": quarter_end,
                        "official": None if pd.isna(official_num) else float(official_num),
                        "fitted": None if pd.isna(fitted_num) else float(fitted_num),
                        "best_forward": None if pd.isna(best_forward_num) else float(best_forward_num),
                        "realized": None if pd.isna(selected_realized_num) else float(selected_realized_num),
                        "official_formula": "",
                        "fitted_formula": "",
                        "best_forward_formula": "",
                        "_preview_priority": -1,
                    }
            preview_priority = {
                "prior_quarter": 0,
                "quarter_open": 1,
                "current_qtd": 2,
                "next_quarter_thesis": 3,
            }
            preview_quarter_specs = [
                ("prior_quarter", prior_market_display_quarter, 2),
                ("quarter_open", quarter_open_display_quarter, 4),
                ("current_qtd", current_market_display_quarter, 6),
                ("next_quarter_thesis", next_thesis_quarter_end, 8),
            ]
            current_chart_model_key = str(current_overlay_model_key or "").strip()
            best_forward_chart_model_key = str(best_forward_overlay_model_key or current_chart_model_key).strip()
            for frame_key, target_quarter_end, source_col in preview_quarter_specs:
                if not isinstance(target_quarter_end, date):
                    continue
                existing_rec = quarterly_chart_rows_by_end.get(target_quarter_end)
                if existing_rec is None:
                    existing_rec = {
                        "quarter_end": target_quarter_end,
                        "official": None,
                        "fitted": None,
                        "best_forward": None,
                        "realized": None,
                        "official_formula": "",
                        "fitted_formula": "",
                        "best_forward_formula": "",
                        "_preview_priority": -1,
                    }
                    quarterly_chart_rows_by_end[target_quarter_end] = existing_rec
                if target_quarter_end in historical_quarter_ends:
                    continue
                if int(existing_rec.get("_preview_priority") or -1) > int(preview_priority.get(frame_key, 0)):
                    continue
                existing_rec["_preview_priority"] = int(preview_priority.get(frame_key, 0))
                existing_rec["official"] = _gpre_preview_frame_value("official_frames", frame_key)
                existing_rec["fitted"] = _gpre_model_preview_frame_value(current_chart_model_key, frame_key)
                existing_rec["best_forward"] = _gpre_model_preview_frame_value(best_forward_chart_model_key, frame_key)
                if official_proxy_comp_row > 0:
                    existing_rec["official_formula"] = f"={get_column_letter(source_col)}{official_proxy_comp_row}"
                if fitted_proxy_comp_row > 0:
                    existing_rec["fitted_formula"] = f"={get_column_letter(source_col)}{fitted_proxy_comp_row}"
                if best_forward_proxy_comp_row > 0:
                    existing_rec["best_forward_formula"] = f"={get_column_letter(source_col)}{best_forward_proxy_comp_row}"
            quarterly_chart_rows = [
                {
                    key: value
                    for key, value in rec.items()
                    if key != "_preview_priority"
                }
                for _, rec in sorted(quarterly_chart_rows_by_end.items(), key=lambda item: item[0])
            ]
            if quarterly_chart_rows:
                quarterly_floor_quarter = date(2023, 3, 31)
                quarterly_chart_rows = [
                    rec for rec in quarterly_chart_rows
                    if isinstance(rec.get("quarter_end"), date) and rec.get("quarter_end") >= quarterly_floor_quarter
                ]
                if len(quarterly_chart_rows) > overlay_quarter_chart_max_points:
                    quarterly_chart_rows = quarterly_chart_rows[-overlay_quarter_chart_max_points:]
            if len(quarterly_chart_rows) >= 2:
                quarterly_chart_title_row = chart_end_row + 2
                while bool(ws.row_dimensions[quarterly_chart_title_row].hidden):
                    quarterly_chart_title_row += 1
                quarterly_chart_anchor_row = quarterly_chart_title_row + 1
                quarterly_chart_end_row = quarterly_chart_anchor_row + overlay_chart_row_span
                quarterly_col_label = 45  # AS
                quarterly_col_official = 46  # AT
                quarterly_col_fitted = 47  # AU
                quarterly_col_best_forward = 48  # AV
                quarterly_col_realized = 49  # AW
                for helper_col in (
                    quarterly_col_label,
                    quarterly_col_official,
                    quarterly_col_fitted,
                    quarterly_col_best_forward,
                    quarterly_col_realized,
                ):
                    ws.column_dimensions[get_column_letter(helper_col)].hidden = True
                quarterly_helper_start_row = quarterly_chart_title_row
                ws.cell(row=quarterly_helper_start_row, column=quarterly_col_label, value="Quarter")
                ws.cell(row=quarterly_helper_start_row, column=quarterly_col_official, value="Approximate market crush ($/gal)")
                ws.cell(row=quarterly_helper_start_row, column=quarterly_col_fitted, value="GPRE crush proxy ($/gal)")
                ws.cell(row=quarterly_helper_start_row, column=quarterly_col_best_forward, value="Best forward lens ($/gal)")
                ws.cell(row=quarterly_helper_start_row, column=quarterly_col_realized, value="Realized GPRE crush margin ($/gal)")
                quarterly_helper_last_row = quarterly_helper_start_row
                quarterly_y_values: List[float] = []
                for helper_row, rec in enumerate(quarterly_chart_rows, start=quarterly_helper_start_row + 1):
                    quarter_end = rec["quarter_end"]
                    quarter_label = (
                        f"{quarter_end.year}-Q{((quarter_end.month - 1) // 3) + 1}"
                        if isinstance(quarter_end, date)
                        else ""
                    )
                    ws.cell(row=helper_row, column=quarterly_col_label, value=quarter_label)
                    official_num = pd.to_numeric(rec.get("official"), errors="coerce")
                    fitted_num = pd.to_numeric(rec.get("fitted"), errors="coerce")
                    best_forward_num = pd.to_numeric(rec.get("best_forward"), errors="coerce")
                    realized_num = pd.to_numeric(rec.get("realized"), errors="coerce")
                    official_formula = str(rec.get("official_formula") or "").strip()
                    fitted_formula = str(rec.get("fitted_formula") or "").strip()
                    best_forward_formula = str(rec.get("best_forward_formula") or "").strip()
                    if official_formula:
                        ws.cell(row=helper_row, column=quarterly_col_official, value=official_formula)
                        ws.cell(row=helper_row, column=quarterly_col_official).number_format = "#,##0.000"
                    elif pd.notna(official_num):
                        ws.cell(row=helper_row, column=quarterly_col_official, value=float(official_num))
                        ws.cell(row=helper_row, column=quarterly_col_official).number_format = "#,##0.000"
                    if pd.notna(official_num):
                        quarterly_y_values.append(float(official_num))
                    if fitted_formula:
                        ws.cell(row=helper_row, column=quarterly_col_fitted, value=fitted_formula)
                        ws.cell(row=helper_row, column=quarterly_col_fitted).number_format = "#,##0.000"
                    elif pd.notna(fitted_num):
                        ws.cell(row=helper_row, column=quarterly_col_fitted, value=float(fitted_num))
                        ws.cell(row=helper_row, column=quarterly_col_fitted).number_format = "#,##0.000"
                    if pd.notna(fitted_num):
                        quarterly_y_values.append(float(fitted_num))
                    if best_forward_formula:
                        ws.cell(row=helper_row, column=quarterly_col_best_forward, value=best_forward_formula)
                        ws.cell(row=helper_row, column=quarterly_col_best_forward).number_format = "#,##0.000"
                    elif pd.notna(best_forward_num):
                        ws.cell(row=helper_row, column=quarterly_col_best_forward, value=float(best_forward_num))
                        ws.cell(row=helper_row, column=quarterly_col_best_forward).number_format = "#,##0.000"
                    if pd.notna(best_forward_num):
                        quarterly_y_values.append(float(best_forward_num))
                    if pd.notna(realized_num):
                        ws.cell(row=helper_row, column=quarterly_col_realized, value=float(realized_num))
                        ws.cell(row=helper_row, column=quarterly_col_realized).number_format = "#,##0.000"
                        quarterly_y_values.append(float(realized_num))
                    quarterly_helper_last_row = helper_row
                ws.merge_cells(start_row=quarterly_chart_title_row, start_column=chart_start_col, end_row=quarterly_chart_title_row, end_column=chart_end_col)
                quarterly_title_cell = ws.cell(
                    row=quarterly_chart_title_row,
                    column=2,
                    value="Approximate market crush, fitted models, and real GPRE crush margin (quarterly)",
                )
                quarterly_title_cell.fill = title_fill
                quarterly_title_cell.font = title_font
                quarterly_title_cell.alignment = align_center
                quarterly_title_cell.border = thin_border
                for cc in range(chart_start_col, chart_end_col + 1):
                    ws.cell(row=quarterly_chart_title_row, column=cc).fill = title_fill
                    ws.cell(row=quarterly_chart_title_row, column=cc).font = title_font
                    ws.cell(row=quarterly_chart_title_row, column=cc).alignment = align_center
                    ws.cell(row=quarterly_chart_title_row, column=cc).border = thin_border
                ws.row_dimensions[quarterly_chart_title_row].height = 18.0

                quarterly_chart = LineChart()
                quarterly_chart.title = None
                quarterly_chart.height = overlay_chart_height
                quarterly_chart.width = overlay_chart_width
                quarterly_chart.x_axis.title = None
                quarterly_chart.x_axis.tickLblPos = "low"
                quarterly_chart.x_axis.majorTickMark = "out"
                quarterly_chart.x_axis.minorTickMark = "none"
                quarterly_chart.x_axis.spPr = GraphicalProperties(
                    ln=LineProperties(noFill=True),
                )
                quarterly_chart.y_axis.title = None
                quarterly_chart.y_axis.axPos = "l"
                quarterly_chart.y_axis.delete = False
                quarterly_chart.y_axis.number_format = "$0.00"
                quarterly_chart.y_axis.tickLblPos = "nextTo"
                quarterly_chart.y_axis.crosses = "min"
                quarterly_chart.y_axis.majorTickMark = "out"
                quarterly_chart.y_axis.minorTickMark = "none"
                quarterly_chart.visible_cells_only = False
                try:
                    quarterly_chart.legend.position = "t"
                    quarterly_chart.legend.overlay = True
                except Exception:
                    pass

                quarterly_data_ref = Reference(
                    ws,
                    min_col=quarterly_col_official,
                    min_row=quarterly_helper_start_row,
                    max_col=quarterly_col_realized,
                    max_row=quarterly_helper_last_row,
                )
                quarterly_chart.add_data(quarterly_data_ref, titles_from_data=True, from_rows=False)
                _apply_chart_text_categories(
                    quarterly_chart,
                    sheet_name=ws.title,
                    col_idx=quarterly_col_label,
                    start_row=quarterly_helper_start_row + 1,
                    end_row=quarterly_helper_last_row,
                )
                try:
                    quarterly_chart.series[0].graphicalProperties.line.solidFill = "2F80ED"
                    quarterly_chart.series[0].graphicalProperties.line.width = 12700
                    quarterly_chart.series[0].marker.symbol = "circle"
                    quarterly_chart.series[0].marker.size = 6
                except Exception:
                    pass
                try:
                    quarterly_chart.series[1].graphicalProperties.line.solidFill = "E67E22"
                    quarterly_chart.series[1].graphicalProperties.line.width = 12700
                    quarterly_chart.series[1].marker.symbol = "diamond"
                    quarterly_chart.series[1].marker.size = 7
                except Exception:
                    pass
                try:
                    quarterly_chart.series[2].graphicalProperties.line.solidFill = "2A9D8F"
                    quarterly_chart.series[2].graphicalProperties.line.width = 12700
                    quarterly_chart.series[2].marker.symbol = "triangle"
                    quarterly_chart.series[2].marker.size = 7
                except Exception:
                    pass
                try:
                    quarterly_chart.series[3].graphicalProperties.line.solidFill = "36454F"
                    quarterly_chart.series[3].graphicalProperties.line.width = 19050
                    quarterly_chart.series[3].marker.symbol = "square"
                    quarterly_chart.series[3].marker.size = 7
                except Exception:
                    pass
                for series_idx in range(min(len(quarterly_chart.series), 4)):
                    try:
                        quarterly_chart.series[series_idx].dLbls = DataLabelList(
                            showLegendKey=False,
                            showVal=True,
                            showSerName=False,
                            showCatName=False,
                            showLeaderLines=False,
                            dLblPos="r",
                            numFmt="#,##0.000",
                        )
                    except Exception:
                        pass

                if quarterly_y_values:
                    quarterly_y_min = min(quarterly_y_values)
                    quarterly_y_max = max(quarterly_y_values)
                else:
                    quarterly_y_min, quarterly_y_max = (-0.50, 0.50)
                quarterly_span = max(float(quarterly_y_max) - float(quarterly_y_min), 0.10)
                quarterly_pad = max(quarterly_span * 0.08, 0.03)
                quarterly_chart_y_min = float(quarterly_y_min) - quarterly_pad
                quarterly_chart_y_max = float(quarterly_y_max) + quarterly_pad
                try:
                    quarterly_chart.y_axis.scaling.min = float(quarterly_chart_y_min)
                    quarterly_chart.y_axis.scaling.max = float(quarterly_chart_y_max)
                except Exception:
                    pass

                quarterly_chart.anchor = TwoCellAnchor(
                    _from=AnchorMarker(col=1, row=quarterly_chart_anchor_row - 1),
                    to=AnchorMarker(col=chart_end_col, row=quarterly_chart_end_row),
                )
                ws.add_chart(quarterly_chart)
                row_idx = max(row_idx, quarterly_chart_end_row + 1)
    _record_writer_substage("write_excel.drivers.render.economics_overlay.charts_helpers", overlay_charts_started)
    return EconomicsOverlayChartWriterResult(row_idx=row_idx)
