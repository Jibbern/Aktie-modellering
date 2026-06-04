"""GPRE commercial/commentary sections for the Economics_Overlay sheet."""
from __future__ import annotations

import re
import time
from copy import copy
from dataclasses import dataclass
from datetime import date
from typing import Any, Callable, Dict, List, Mapping, Sequence, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border


@dataclass(frozen=True)
class GpreEconomicsOverlayCommercialDeps:
    ws: Any
    is_gpre_profile: bool
    row_idx: int
    gpre_commercial_setup_rows: Sequence[Mapping[str, Any]]
    derivative_bridge_by_quarter: Mapping[date, Mapping[str, Any]]
    overlay_gpre_end_col: int
    analysis_theme: Mapping[str, Any]
    body_font: Any
    bold_font: Any
    horizon_font: Any
    setup_font: Any
    row_border: Any
    thin_border: Any
    quarter_separator_side: Any
    overlay_commentary_section_row_height: float
    overlay_commentary_header_row_height: float
    overlay_commentary_year_band_row_height: float
    overlay_commercial_section_row_height: float
    overlay_commercial_header_row_height: float
    overlay_commercial_year_band_row_height: float
    overlay_commercial_row_max_height: float
    overlay_support_row_height: float
    add_comment: Callable[[str, Any], None]
    ensure_terminal_period: Callable[[Any], str]
    estimate_wrapped_row_height: Callable[..., float]
    normalize_text: Callable[[Any], str]
    overlay_driver_source_priority: Callable[[Any], int]
    record_writer_substage: Callable[[str, float], None]
    write_header_row: Callable[..., int]
    write_section_bar: Callable[..., int]
    write_year_band: Callable[..., int]


@dataclass(frozen=True)
class GpreEconomicsOverlayCommercialResult:
    row_idx: int
    management_row_count: int
    commercial_row_count: int


def _format_derivative_usd_short(usd_value: Any) -> str:
    val = pd.to_numeric(usd_value, errors="coerce")
    if pd.isna(val):
        return ""
    sign = "-" if float(val) < 0 else ""
    return f"{sign}${abs(float(val)) / 1_000_000.0:,.1f}m"


def _with_top_separator(border_in: Any, quarter_separator_side: Any) -> Border:
    border_obj = border_in if isinstance(border_in, Border) else Border()
    return Border(
        left=copy(border_obj.left),
        right=copy(border_obj.right),
        top=copy(quarter_separator_side),
        bottom=copy(border_obj.bottom),
        diagonal=copy(border_obj.diagonal),
        diagonalUp=bool(border_obj.diagonalUp),
        diagonalDown=bool(border_obj.diagonalDown),
        outline=bool(border_obj.outline),
        vertical=copy(border_obj.vertical),
        horizontal=copy(border_obj.horizontal),
    )


def _commentary_quarter_separator_needed(
    previous_quarter_label: str,
    previous_year_band: str,
    current_quarter_label: str,
    current_year_band: str,
) -> bool:
    prev_q = str(previous_quarter_label or "").strip()
    curr_q = str(current_quarter_label or "").strip()
    prev_y = str(previous_year_band or "").strip()
    curr_y = str(current_year_band or "").strip()
    return bool(prev_q and curr_q and prev_y and curr_y and prev_y == curr_y and prev_q != curr_q)


def _year_band_label(rec: Mapping[str, Any]) -> str:
    horizon_norm = str(rec.get("horizon_period_norm") or "").strip()
    match = re.match(r"Q(20\d{2})Q([1-4])$", horizon_norm)
    if match:
        return "2026 / current" if int(match.group(1)) >= 2026 else str(int(match.group(1)))
    horizon_lbl = str(rec.get("horizon_quarter") or "").strip()
    if re.search(r"\b2026\b", horizon_lbl):
        return "2026 / current"
    src_q = rec.get("source_quarter")
    if isinstance(src_q, date):
        return "2026 / current" if int(src_q.year) >= 2026 else str(int(src_q.year))
    return "Other"


def _year_band_sort_rank(label_in: Any) -> int:
    label_txt = str(label_in or "").strip()
    return (
        0 if label_txt == "2026 / current"
        else 1 if label_txt == "2025"
        else 2 if label_txt == "2024"
        else 3 if label_txt == "2023"
        else 99
    )


def _gpre_management_commentary_rows(deps: GpreEconomicsOverlayCommercialDeps) -> List[Dict[str, Any]]:
    if not (deps.is_gpre_profile and deps.gpre_commercial_setup_rows):
        return []
    out_rows: List[Dict[str, Any]] = []
    seen_keys: set[Tuple[str, str]] = set()
    commentary_source_rows = sorted(
        [
            dict(rec)
            for rec in deps.gpre_commercial_setup_rows
            if bool(rec.get("show_in_management_commentary", True))
            and str(rec.get("commentary_home") or "overlay_management") == "overlay_management"
        ],
        key=lambda rec: (
            -int(pd.to_datetime(rec.get("source_quarter"), errors="coerce").strftime("%Y%m%d"))
            if not pd.isna(pd.to_datetime(rec.get("source_quarter"), errors="coerce"))
            else 0,
            int(rec.get("commentary_priority") or 50),
            deps.overlay_driver_source_priority(rec.get("source_type")),
            str(rec.get("setup_display") or ""),
        ),
    )
    for rec in commentary_source_rows:
        commentary_txt = deps.ensure_terminal_period(deps.normalize_text(str(rec.get("commentary_text") or "")))
        if not commentary_txt:
            continue
        key = (
            str(rec.get("source_quarter_label") or ""),
            commentary_txt.lower(),
        )
        if key in seen_keys:
            continue
        seen_keys.add(key)
        out_rows.append(
            {
                "horizon_quarter": str(rec.get("horizon_quarter") or ""),
                "source_quarter_label": str(rec.get("source_quarter_label") or ""),
                "source_quarter": rec.get("source_quarter"),
                "commentary_text": commentary_txt,
                "_display_order": len(out_rows),
                "comment_text": "\n".join(
                    [
                        part
                        for part in (
                            f"Source: {str(rec.get('source_type') or '').strip().title()} | Confidence: {str(rec.get('confidence') or '').strip().title()}",
                            f"Location: {str(rec.get('source_location') or '').strip()}",
                            f"Excerpt: {str(rec.get('source_excerpt') or '').strip()}",
                        )
                        if part and not part.endswith(": ")
                    ]
                ),
            }
        )
    if deps.derivative_bridge_by_quarter:
        latest_der_q = max(deps.derivative_bridge_by_quarter)
        latest_der = dict(deps.derivative_bridge_by_quarter.get(latest_der_q) or {})
        der_source = str(latest_der.get("derivative_source_document") or "").strip()
        der_note = str(latest_der.get("derivative_notes") or "").strip()
        q_label = f"{latest_der_q.year}-Q{((latest_der_q.month - 1) // 3) + 1}"
        pnl_val = pd.to_numeric(latest_der.get("derivative_gain_loss_pnl_total_usd"), errors="coerce")
        rev_val = pd.to_numeric(latest_der.get("derivative_gain_loss_revenue_usd"), errors="coerce")
        cogs_val = pd.to_numeric(latest_der.get("derivative_gain_loss_cogs_usd"), errors="coerce")
        if pd.notna(pnl_val):
            text = (
                f"Derivative P&L impact is already embedded in reported revenue/COGS "
                f"({_format_derivative_usd_short(pnl_val)} total; revenue {_format_derivative_usd_short(rev_val)}, "
                f"COGS {_format_derivative_usd_short(cogs_val)})."
            )
            if ("Derivative / OCI", text.lower()) not in seen_keys:
                seen_keys.add(("Derivative / OCI", text.lower()))
                out_rows.append(
                    {
                        "horizon_quarter": "Actual",
                        "source_quarter_label": q_label,
                        "source_quarter": latest_der_q,
                        "commentary_text": text,
                        "_display_order": len(out_rows),
                        "comment_text": "\n".join(part for part in (f"Source: 10-Q/10-K derivative footnote", f"Document: {der_source}", der_note) if part),
                    }
                )
        oci_val = pd.to_numeric(latest_der.get("derivative_oci_current_period_usd"), errors="coerce")
        if pd.notna(oci_val):
            text = f"Derivative OCI movement was {_format_derivative_usd_short(oci_val)}; (unrealized hedge cash-flow)."
            if ("Derivative / OCI", text.lower()) not in seen_keys:
                seen_keys.add(("Derivative / OCI", text.lower()))
                out_rows.append(
                    {
                        "horizon_quarter": "Actual",
                        "source_quarter_label": q_label,
                        "source_quarter": latest_der_q,
                        "commentary_text": text,
                        "_display_order": len(out_rows),
                        "comment_text": "\n".join(part for part in (f"Source: 10-Q/10-K comprehensive income footnote", f"Document: {der_source}", der_note) if part),
                    }
                )
    out_rows.sort(
        key=lambda rec: (
            _year_band_sort_rank(_year_band_label(rec)),
            -int(pd.to_datetime(rec.get("source_quarter"), errors="coerce").strftime("%Y%m%d"))
            if not pd.isna(pd.to_datetime(rec.get("source_quarter"), errors="coerce"))
            else 0,
            int(rec.get("_display_order") or 0),
        )
    )
    for rec in out_rows:
        rec.pop("_display_order", None)
    return out_rows


def _write_gpre_management_commentary_section(
    deps: GpreEconomicsOverlayCommercialDeps,
    row_num: int,
    management_rows: Sequence[Mapping[str, Any]],
) -> int:
    if not (deps.is_gpre_profile and management_rows):
        return row_num
    ws = deps.ws
    row_num = deps.write_section_bar(
        row_num,
        "Management commentary",
        end_col=deps.overlay_gpre_end_col,
        primary=True,
        row_height=deps.overlay_commentary_section_row_height,
    )
    row_num = deps.write_header_row(
        row_num,
        ["Horizon", "Stated in", "Commentary"],
        spans=[
            (1, 1, "Horizon"),
            (2, 2, "Stated in"),
            (3, 17, "Commentary"),
        ],
        row_height=deps.overlay_commentary_header_row_height,
    )
    last_year_band = ""
    last_stated_in = ""
    data_fill = copy(deps.analysis_theme["neutral_fill_alt"])
    for rec in management_rows:
        year_band = _year_band_label(rec)
        if year_band != last_year_band:
            row_num = deps.write_year_band(
                row_num,
                year_band,
                end_col=deps.overlay_gpre_end_col,
                row_height=deps.overlay_commentary_year_band_row_height,
            )
            last_year_band = year_band
            last_stated_in = ""
        stated_in_txt = str(rec.get("source_quarter_label") or "")
        add_quarter_separator = _commentary_quarter_separator_needed(
            last_stated_in,
            last_year_band,
            stated_in_txt,
            year_band,
        )
        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=deps.overlay_gpre_end_col)
        ws.cell(row=row_num, column=1, value=str(rec.get("horizon_quarter") or ""))
        ws.cell(row=row_num, column=2, value=stated_in_txt)
        ws.cell(row=row_num, column=3, value=str(rec.get("commentary_text") or ""))
        for cc in range(1, deps.overlay_gpre_end_col + 1):
            wrap_cols = {1, 2, 3}
            ws.cell(row=row_num, column=cc).fill = copy(data_fill)
            ws.cell(row=row_num, column=cc).border = Border()
            ws.cell(row=row_num, column=cc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=cc in wrap_cols)
            ws.cell(row=row_num, column=cc).font = deps.body_font
        ws.cell(row=row_num, column=1).font = deps.horizon_font
        ws.cell(row=row_num, column=2).font = deps.horizon_font
        if add_quarter_separator:
            for cc in range(1, deps.overlay_gpre_end_col + 1):
                ws.cell(row=row_num, column=cc).border = _with_top_separator(
                    ws.cell(row=row_num, column=cc).border,
                    deps.quarter_separator_side,
                )
        if str(rec.get("comment_text") or "").strip():
            deps.add_comment(f"C{row_num}", str(rec.get("comment_text") or "").strip())
        ws.row_dimensions[row_num].height = 19.5
        last_stated_in = stated_in_txt or last_stated_in
        row_num += 1
    return row_num


def _write_gpre_commercial_setup_section(
    deps: GpreEconomicsOverlayCommercialDeps,
    row_num: int,
) -> Tuple[int, int]:
    if not (deps.is_gpre_profile and deps.gpre_commercial_setup_rows):
        return row_num, 0
    ws = deps.ws
    visible_setup_rows = [dict(it) for it in reversed(deps.gpre_commercial_setup_rows) if bool(it.get("show_in_setup", True))]

    def _commercial_row_height(rec_in: Mapping[str, Any]) -> float:
        coverage_txt_local = str(rec_in.get("coverage_text") or "").strip()
        locked_bits_local = [
            str(rec_in.get("locked_margin_text") or "").strip(),
            str(rec_in.get("legs_involved") or "").strip(),
        ]
        locked_txt_local = " | ".join([x for x in locked_bits_local if x])
        coverage_width = sum(float(ws.column_dimensions[col].width or 15.0) for col in ("E", "F", "G"))
        locked_width = sum(float(ws.column_dimensions[col].width or 15.0) for col in ("H", "I", "J", "K"))
        effect_width = sum(float(ws.column_dimensions[col].width or 15.0) for col in ("L", "M", "N"))
        takeaway_width = sum(float(ws.column_dimensions[col].width or 15.0) for col in ("O", "P", "Q"))
        narrative_texts_local = [
            coverage_txt_local,
            locked_txt_local,
            str(rec_in.get("result_effect") or "").strip(),
            str(rec_in.get("management_takeaway") or "").strip(),
        ]
        long_field_count_local = sum(1 for txt in narrative_texts_local if len(str(txt or "")) >= 70)
        max_text_len_local = max((len(str(txt or "")) for txt in narrative_texts_local), default=0)
        dense_row_floor_local = 51.0 if max_text_len_local >= 90 or long_field_count_local >= 2 else deps.overlay_support_row_height
        estimated_height = max(
            deps.estimate_wrapped_row_height(coverage_txt_local, coverage_width, 19, 11, min_lines=1, max_lines=5),
            deps.estimate_wrapped_row_height(locked_txt_local, locked_width, 19, 11, min_lines=1, max_lines=5),
            deps.estimate_wrapped_row_height(str(rec_in.get("result_effect") or "").strip(), effect_width, 19, 11, min_lines=1, max_lines=4),
            deps.estimate_wrapped_row_height(str(rec_in.get("management_takeaway") or "").strip(), takeaway_width, 19, 11, min_lines=1, max_lines=5),
        )
        if estimated_height >= 29.0:
            estimated_height += 2.0
        if estimated_height >= 45.0:
            estimated_height += 2.5
        return max(
            dense_row_floor_local,
            min(
                deps.overlay_commercial_row_max_height,
                estimated_height,
            ),
        )

    row_num = deps.write_section_bar(
        row_num,
        "Commercial / hedge setup",
        end_col=deps.overlay_gpre_end_col,
        primary=True,
        row_height=deps.overlay_commercial_section_row_height,
    )
    row_num = deps.write_header_row(
        row_num,
        [
            "Horizon",
            "Stated in",
            "Setup",
            "Coverage / openness",
            "Locked margin / legs",
            "Effect on results",
            "Takeaway",
        ],
        spans=[
            (1, 1, "Horizon"),
            (2, 2, "Stated in"),
            (3, 4, "Setup"),
            (5, 7, "Coverage / openness"),
            (8, 11, "Locked margin / legs"),
            (12, 14, "Effect on results"),
            (15, 17, "Takeaway"),
        ],
        row_height=deps.overlay_commercial_header_row_height,
    )
    last_year_band = ""
    last_stated_in = ""
    data_fill = copy(deps.analysis_theme["neutral_fill_alt"])
    for rec_idx, rec in enumerate(visible_setup_rows):
        year_band = _year_band_label(rec)
        if year_band != last_year_band:
            row_num = deps.write_year_band(
                row_num,
                year_band,
                end_col=deps.overlay_gpre_end_col,
                row_height=deps.overlay_commercial_year_band_row_height,
            )
            last_year_band = year_band
            last_stated_in = ""
        stated_in_txt = str(rec.get("source_quarter_label") or "")
        add_quarter_separator = bool(last_stated_in and stated_in_txt and stated_in_txt != last_stated_in)
        coverage_txt = str(rec.get("coverage_text") or "").strip()
        locked_bits = [
            str(rec.get("locked_margin_text") or "").strip(),
            str(rec.get("legs_involved") or "").strip(),
        ]
        locked_txt = " | ".join([x for x in locked_bits if x])
        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
        ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=7)
        ws.merge_cells(start_row=row_num, start_column=8, end_row=row_num, end_column=11)
        ws.merge_cells(start_row=row_num, start_column=12, end_row=row_num, end_column=14)
        ws.merge_cells(start_row=row_num, start_column=15, end_row=row_num, end_column=17)
        ws.cell(row=row_num, column=1, value=str(rec.get("horizon_quarter") or ""))
        ws.cell(row=row_num, column=2, value=stated_in_txt)
        ws.cell(row=row_num, column=3, value=str(rec.get("setup_display") or ""))
        ws.cell(row=row_num, column=5, value=coverage_txt)
        ws.cell(row=row_num, column=8, value=locked_txt)
        ws.cell(row=row_num, column=12, value=str(rec.get("result_effect") or ""))
        ws.cell(row=row_num, column=15, value=str(rec.get("management_takeaway") or ""))
        for cc in range(1, deps.overlay_gpre_end_col + 1):
            wrap_cols = {1, 2, 3, 5, 8, 12, 15}
            ws.cell(row=row_num, column=cc).fill = copy(data_fill)
            ws.cell(row=row_num, column=cc).border = deps.row_border
            ws.cell(row=row_num, column=cc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=cc in wrap_cols)
            ws.cell(row=row_num, column=cc).font = deps.body_font
        ws.cell(row=row_num, column=1).font = deps.horizon_font
        ws.cell(row=row_num, column=2).font = deps.horizon_font
        ws.cell(row=row_num, column=3).font = deps.setup_font
        if add_quarter_separator:
            for cc in range(1, deps.overlay_gpre_end_col + 1):
                ws.cell(row=row_num, column=cc).border = _with_top_separator(
                    ws.cell(row=row_num, column=cc).border,
                    deps.quarter_separator_side,
                )
        comment_text = "\n".join(
            [
                part
                for part in (
                    f"Source: {str(rec.get('source_type') or '').strip().title()} | Confidence: {str(rec.get('confidence') or '').strip().title()}",
                    f"Location: {str(rec.get('source_location') or '').strip()}",
                    f"Excerpt: {str(rec.get('source_excerpt') or '').strip()}",
                )
                if part and not part.endswith(": ")
            ]
        )
        if comment_text:
            deps.add_comment(f"C{row_num}", comment_text)
        current_row_height = _commercial_row_height(rec)
        ws.row_dimensions[row_num].height = current_row_height
        last_stated_in = stated_in_txt or last_stated_in
        row_num += 1
        next_rec = visible_setup_rows[rec_idx + 1] if rec_idx + 1 < len(visible_setup_rows) else None
        if next_rec is not None:
            next_year_band = _year_band_label(next_rec)
            if year_band == next_year_band:
                for cc in range(1, deps.overlay_gpre_end_col + 1):
                    ws.cell(row=row_num, column=cc).fill = copy(data_fill)
                ws.row_dimensions[row_num].height = 6.0
                row_num += 1
    return row_num, len(visible_setup_rows)


def write_gpre_economics_overlay_commercial_sections(
    deps: GpreEconomicsOverlayCommercialDeps,
) -> GpreEconomicsOverlayCommercialResult:
    row_num = int(deps.row_idx)
    if not (deps.is_gpre_profile and deps.gpre_commercial_setup_rows):
        return GpreEconomicsOverlayCommercialResult(
            row_idx=row_num,
            management_row_count=0,
            commercial_row_count=0,
        )

    management_rows = _gpre_management_commentary_rows(deps)
    overlay_management_started = time.perf_counter()
    row_num = _write_gpre_management_commentary_section(deps, row_num, management_rows)
    deps.record_writer_substage(
        "write_excel.drivers.render.economics_overlay.management_commentary",
        overlay_management_started,
    )
    row_num = max(row_num + 1, 37)

    overlay_commercial_started = time.perf_counter()
    row_num, commercial_row_count = _write_gpre_commercial_setup_section(deps, row_num)
    deps.record_writer_substage(
        "write_excel.drivers.render.economics_overlay.commercial_setup",
        overlay_commercial_started,
    )
    row_num = max(row_num + 1, 68)

    return GpreEconomicsOverlayCommercialResult(
        row_idx=row_num,
        management_row_count=len(management_rows),
        commercial_row_count=commercial_row_count,
    )
