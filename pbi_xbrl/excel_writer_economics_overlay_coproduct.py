"""GPRE coproduct visible section for the Economics_Overlay sheet."""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date
import time
from typing import Any, Callable, Dict, List, Mapping, Optional, Sequence, Set, Tuple

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class GpreEconomicsOverlayCoproductDeps:
    wb: Any
    ws: Any
    row_idx: int
    is_gpre_profile: bool
    gpre_commercial_setup_rows: Sequence[Mapping[str, Any]]
    gpre_basis_sandbox_layout: Mapping[str, Any]
    gpre_overlay_coproduct_start_row: int | None
    market_rows: Mapping[str, Any]
    coeff_rows: Mapping[str, Any]
    current_qtd_market_snapshot: Mapping[str, Any]
    prior_market_display_quarter: Any
    quarter_open_display_quarter: Any
    current_market_display_quarter: Any
    next_thesis_quarter_end: Any
    prior_market_display_quarter_txt: str
    quarter_open_overlay_header_txt: str
    next_thesis_quarter_txt: str
    overlay_chart_width: float
    overlay_chart_height: float
    overlay_chart_row_span: int
    overlay_quarter_chart_max_points: int
    overlay_header_row_height: float
    title_fill: Any
    title_font: Any
    header_fill: Any
    body_font: Any
    bold_font: Any
    thin_border: Any
    zebra_fill_light: Any
    zebra_fill_dark: Any
    intro_fill: Any
    section_fill: Any
    align_center: Any
    overlay_as_of_header_text: Callable[[Any], str]
    overlay_coefficient_detail: Callable[[str], Mapping[str, Any]]
    parse_quarter_label_text: Callable[[Any], Any]
    quarter_label_short: Callable[[Any], str]
    apply_chart_text_categories: Callable[..., None]
    write_overlay_subheader_row: Callable[..., int]
    record_writer_substage: Callable[[str, float], None]


@dataclass(frozen=True)
class GpreEconomicsOverlayCoproductResult:
    row_idx: int
    row_count: int
    chart_count: int


def write_gpre_economics_overlay_coproduct_section(
    deps: GpreEconomicsOverlayCoproductDeps,
) -> GpreEconomicsOverlayCoproductResult:
    wb = deps.wb
    ws = deps.ws
    row_idx = int(deps.row_idx)
    initial_row_idx = row_idx
    initial_chart_count = len(getattr(ws, "_charts", []) or [])
    is_gpre_profile = deps.is_gpre_profile
    gpre_commercial_setup_rows = deps.gpre_commercial_setup_rows
    gpre_basis_sandbox_layout = deps.gpre_basis_sandbox_layout
    gpre_overlay_coproduct_start_row = deps.gpre_overlay_coproduct_start_row
    market_rows = deps.market_rows
    coeff_rows = deps.coeff_rows
    current_qtd_market_snapshot = deps.current_qtd_market_snapshot
    prior_market_display_quarter = deps.prior_market_display_quarter
    quarter_open_display_quarter = deps.quarter_open_display_quarter
    current_market_display_quarter = deps.current_market_display_quarter
    next_thesis_quarter_end = deps.next_thesis_quarter_end
    prior_market_display_quarter_txt = deps.prior_market_display_quarter_txt
    quarter_open_overlay_header_txt = deps.quarter_open_overlay_header_txt
    next_thesis_quarter_txt = deps.next_thesis_quarter_txt
    overlay_chart_width = deps.overlay_chart_width
    overlay_chart_height = deps.overlay_chart_height
    overlay_chart_row_span = deps.overlay_chart_row_span
    overlay_quarter_chart_max_points = deps.overlay_quarter_chart_max_points
    overlay_header_row_height = deps.overlay_header_row_height
    title_fill = deps.title_fill
    title_font = deps.title_font
    header_fill = deps.header_fill
    body_font = deps.body_font
    bold_font = deps.bold_font
    thin_border = deps.thin_border
    zebra_fill_light = deps.zebra_fill_light
    zebra_fill_dark = deps.zebra_fill_dark
    intro_fill = deps.intro_fill
    section_fill = deps.section_fill
    align_center = deps.align_center
    _overlay_as_of_header_text = deps.overlay_as_of_header_text
    _overlay_coefficient_detail = deps.overlay_coefficient_detail
    _parse_quarter_label_text = deps.parse_quarter_label_text
    _quarter_label_short = deps.quarter_label_short
    _apply_chart_text_categories = deps.apply_chart_text_categories
    _write_overlay_subheader_row = deps.write_overlay_subheader_row
    _record_writer_substage = deps.record_writer_substage

    overlay_coproduct_block_started = time.perf_counter()
    if is_gpre_profile and gpre_commercial_setup_rows and bool(gpre_basis_sandbox_layout.get("coproduct_visible_block_allowed")):
        visible_coproduct_start_row = max(int(gpre_overlay_coproduct_start_row or 176), 176)
        visible_coproduct_section_row = max(int(visible_coproduct_start_row) - 1, 177)
        visible_coproduct_header_row = visible_coproduct_start_row
        visible_coproduct_subheader_row = visible_coproduct_header_row + 1
        visible_coproduct_separator_row = visible_coproduct_header_row + 4
        visible_coproduct_rows = {
            "corn_oil_price": visible_coproduct_header_row + 2,
            "distillers_price": visible_coproduct_header_row + 3,
            "coproduct_credit_per_gal": visible_coproduct_header_row + 5,
            "coproduct_credit_usd_m": visible_coproduct_header_row + 6,
        }
        source_mode_start_col = 11  # K
        source_mode_end_col = 21  # U
        ws.merge_cells(
            start_row=visible_coproduct_section_row,
            start_column=1,
            end_row=visible_coproduct_section_row,
            end_column=21,
        )
        visible_coproduct_section_title = ws.cell(
            row=visible_coproduct_section_row,
            column=1,
            value="Coproducts",
        )
        visible_coproduct_section_title.fill = copy(title_fill)
        visible_coproduct_section_title.font = copy(title_font)
        visible_coproduct_section_title.alignment = align_center
        visible_coproduct_section_title.border = copy(thin_border)
        for cc in range(1, 22):
            ws.cell(row=visible_coproduct_section_row, column=cc).fill = copy(title_fill)
            ws.cell(row=visible_coproduct_section_row, column=cc).font = copy(title_font)
            ws.cell(row=visible_coproduct_section_row, column=cc).alignment = align_center
            ws.cell(row=visible_coproduct_section_row, column=cc).border = copy(thin_border)
        ws.row_dimensions[visible_coproduct_section_row].height = 18.0
        quarter_spans = [
            (2, 3, "Prior quarter"),
            (4, 5, "Quarter-open outlook"),
            (6, 7, "Current QTD"),
            (8, 9, "Next quarter outlook"),
        ]
        for start_col, end_col, header_txt in quarter_spans:
            ws.merge_cells(
                start_row=visible_coproduct_header_row,
                start_column=start_col,
                end_row=visible_coproduct_header_row,
                end_column=end_col,
            )
        ws.merge_cells(
            start_row=visible_coproduct_header_row,
            start_column=source_mode_start_col,
            end_row=visible_coproduct_header_row,
            end_column=source_mode_end_col,
        )
        for cc in range(1, source_mode_end_col + 1):
            ws.cell(row=visible_coproduct_header_row, column=cc).fill = copy(header_fill)
            ws.cell(row=visible_coproduct_header_row, column=cc).font = copy(bold_font)
            ws.cell(row=visible_coproduct_header_row, column=cc).border = copy(thin_border)
            ws.cell(row=visible_coproduct_header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=visible_coproduct_header_row, column=1, value="Coproduct economics")
        for start_col, _end_col, header_txt in quarter_spans:
            ws.cell(row=visible_coproduct_header_row, column=start_col, value=header_txt)
        ws.cell(row=visible_coproduct_header_row, column=10, value="Unit")
        ws.cell(row=visible_coproduct_header_row, column=source_mode_start_col, value="Source mode")
        ws.row_dimensions[visible_coproduct_header_row].height = overlay_header_row_height
        _write_overlay_subheader_row(
            visible_coproduct_subheader_row,
            prior_txt=prior_market_display_quarter_txt,
            quarter_open_txt=quarter_open_overlay_header_txt,
            current_txt=_overlay_as_of_header_text(current_qtd_market_snapshot.get("as_of") if isinstance(current_qtd_market_snapshot, dict) else None),
            thesis_txt=next_thesis_quarter_txt,
            note_start_col=11,
            note_end_col=21,
            row_height=21.0,
        )

        build_up_rows = dict(((gpre_basis_sandbox_layout.get("approx_market_crush_build_up") or {}).get("econ_rows")) or {})
        corn_oil_source_row = int(market_rows.get("renewable_corn_oil_price") or 0)
        distillers_source_row = int(market_rows.get("distillers_grains_price") or 0)
        coproduct_credit_source_row = int(build_up_rows.get("coproduct_credit") or 0)
        ethanol_yield_row = int(coeff_rows.get("ethanol_yield") or 0)
        ethanol_yield_ref = f"$B${ethanol_yield_row}" if ethanol_yield_row > 0 else ""
        coproduct_frame_key_by_col = {
            2: "prior_quarter",
            4: "quarter_open",
            6: "current_qtd",
            8: "next_quarter_thesis",
        }
        coproduct_frame_layout = dict(gpre_basis_sandbox_layout.get("coproduct_frame_summary") or {})
        frame_rows_by_key = dict(coproduct_frame_layout.get("frame_rows") or {})
        frame_value_col_map = {
            "corn_oil_price": int(coproduct_frame_layout.get("renewable_corn_oil_col") or 0),
            "distillers_price": int(coproduct_frame_layout.get("distillers_grains_col") or 0),
            "coproduct_credit_per_gal": int(coproduct_frame_layout.get("approximate_coproduct_credit_per_gal_col") or 0),
            "coproduct_credit_usd_m": int(coproduct_frame_layout.get("approximate_coproduct_credit_usd_m_col") or 0),
        }
        frame_source_mode_col = int(coproduct_frame_layout.get("resolved_source_mode_col") or 0)
        frame_coverage_col = int(coproduct_frame_layout.get("coverage_col") or 0)
        build_up_frame_col_by_key = {
            "prior_quarter": 3,
            "quarter_open": 5,
            "current_qtd": 7,
            "next_quarter_thesis": 9,
        }
        build_up_corn_oil_contribution_row = int(build_up_rows.get("corn_oil_contribution") or 0)

        visible_coproduct_specs = [
            {
                "row": visible_coproduct_rows["corn_oil_price"],
                "label": "Renewable corn oil price",
                "frame_value_key": "corn_oil_price",
                "unit": "$/lb",
                "source_formula": "Weighted active-capacity quarterly resolver; NWER first, AMS 3618 second. See sandbox for source mode, coverage, and frozen next-quarter rule.",
            },
            {
                "row": visible_coproduct_rows["distillers_price"],
                "label": "Distillers grains price",
                "frame_value_key": "distillers_price",
                "unit": "$/lb",
                "source_formula": "Weighted active-capacity quarterly resolver; NWER first, AMS 3618 second. See sandbox for source mode, coverage, and frozen next-quarter rule.",
            },
            {
                "row": visible_coproduct_rows["coproduct_credit_per_gal"],
                "label": "Approximate coproduct credit ($/gal)",
                "frame_value_key": "coproduct_credit_per_gal",
                "unit": "$/gal",
                "source_formula": "Weighted sandbox build-up divided by ethanol yield; See sandbox for source mode, coverage, and frozen next-quarter rule.",
                "row_height": 28.0,
            },
            {
                "row": visible_coproduct_rows["coproduct_credit_usd_m"],
                "label": "Approximate coproduct credit ($m)",
                "frame_value_key": "coproduct_credit_usd_m",
                "unit": "$m",
                "source_formula": "Derived from weighted coproduct credit ($/gal) times the frame-specific implied gallons basis already used by the crush $m rows.",
                "row_height": 28.0,
            },
        ]

        def _coproduct_frame_formula(frame_key: str, frame_value_key: str) -> str:
            frame_row_num = int(frame_rows_by_key.get(str(frame_key or "")) or 0)
            frame_col_num = int(frame_value_col_map.get(str(frame_value_key or "")) or 0)
            if frame_row_num <= 0 or frame_col_num <= 0:
                return '=""'
            return f'=IF(ISNUMBER(Basis_Proxy_Sandbox!${get_column_letter(frame_col_num)}${frame_row_num}),Basis_Proxy_Sandbox!${get_column_letter(frame_col_num)}${frame_row_num},"")'

        basis_proxy_ws_for_coproduct = wb["Basis_Proxy_Sandbox"] if "Basis_Proxy_Sandbox" in wb.sheetnames else None

        def _coproduct_frame_value(frame_key: str, frame_value_key: str) -> Any:
            frame_row_num = int(frame_rows_by_key.get(str(frame_key or "")) or 0)
            frame_col_num = int(frame_value_col_map.get(str(frame_value_key or "")) or 0)
            if frame_row_num <= 0 or frame_col_num <= 0 or basis_proxy_ws_for_coproduct is None:
                return _coproduct_frame_formula(frame_key, frame_value_key)
            val = basis_proxy_ws_for_coproduct.cell(row=frame_row_num, column=frame_col_num).value
            val_num = pd.to_numeric(val, errors="coerce")
            if pd.notna(val_num):
                return float(val_num)
            return ""

        for spec_idx, spec in enumerate(visible_coproduct_specs, start=1):
            row_num = int(spec["row"])
            row_fill = copy(zebra_fill_light if (spec_idx % 2 == 1) else zebra_fill_dark)
            for start_col, end_col, _header_txt in quarter_spans:
                ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
            ws.merge_cells(
                start_row=row_num,
                start_column=source_mode_start_col,
                end_row=row_num,
                end_column=source_mode_end_col,
            )
            for cc in range(1, source_mode_end_col + 1):
                ws.cell(row=row_num, column=cc).fill = copy(row_fill)
                ws.cell(row=row_num, column=cc).font = copy(body_font)
                ws.cell(row=row_num, column=cc).border = copy(thin_border)
                ws.cell(row=row_num, column=cc).alignment = (
                    Alignment(horizontal="left", vertical="center", wrap_text=True)
                    if cc in {1, source_mode_start_col}
                    else align_center
                )
            ws.cell(row=row_num, column=1, value=str(spec["label"] or ""))
            for target_col, frame_key in coproduct_frame_key_by_col.items():
                value_cell = ws.cell(
                    row=row_num,
                    column=target_col,
                    value=_coproduct_frame_value(frame_key, str(spec.get("frame_value_key") or "")),
                )
                value_cell.number_format = "#,##0.0" if str(spec.get("unit") or "").strip() == "$m" else "#,##0.000"
            ws.cell(row=row_num, column=10, value=str(spec["unit"] or ""))
            ws.cell(row=row_num, column=source_mode_start_col, value=str(spec["source_formula"] or ""))
            ws.row_dimensions[row_num].height = float(spec.get("row_height") or (24.0 if spec_idx < 3 else 28.0))
        for cc in range(1, source_mode_end_col + 1):
            separator_cell = ws.cell(row=visible_coproduct_separator_row, column=cc, value="")
            separator_cell.fill = copy(intro_fill)
            separator_cell.border = copy(thin_border)
            separator_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[visible_coproduct_separator_row].height = 8.0
        row_idx = max(row_idx, visible_coproduct_rows["coproduct_credit_usd_m"] + 2)

        coproduct_history_layout = dict(gpre_basis_sandbox_layout.get("coproduct_quarterly_history") or {})
        basis_proxy_ws = wb["Basis_Proxy_Sandbox"] if "Basis_Proxy_Sandbox" in wb.sheetnames else None
        history_first_row = int(coproduct_history_layout.get("first_data_row") or 0)
        history_last_row = int(coproduct_history_layout.get("last_data_row") or 0)
        history_quarter_col = int(coproduct_history_layout.get("quarter_col") or 0)
        history_credit_per_gal_col = int(coproduct_history_layout.get("approximate_coproduct_credit_per_gal_col") or 0)
        history_credit_usd_m_col = int(coproduct_history_layout.get("approximate_coproduct_credit_usd_m_col") or 0)
        history_source_mode_col = int(coproduct_history_layout.get("resolved_source_mode_col") or 0)
        history_coverage_col = int(coproduct_history_layout.get("coverage_col") or 0)
        frame_priority_for_history = {
            "prior_quarter": 0,
            "quarter_open": 1,
            "current_qtd": 2,
            "next_quarter_thesis": 3,
        }
        frame_key_by_history_quarter: Dict[date, str] = {}
        for frame_key, target_quarter_end in (
            ("prior_quarter", prior_market_display_quarter),
            ("quarter_open", quarter_open_display_quarter),
            ("current_qtd", current_market_display_quarter),
            ("next_quarter_thesis", next_thesis_quarter_end),
        ):
            if not isinstance(target_quarter_end, date):
                continue
            existing_frame_key = frame_key_by_history_quarter.get(target_quarter_end)
            if existing_frame_key is not None and int(frame_priority_for_history.get(existing_frame_key, -1)) > int(frame_priority_for_history.get(frame_key, -1)):
                continue
            frame_key_by_history_quarter[target_quarter_end] = frame_key

        history_record_by_quarter: Dict[date, Dict[str, Any]] = {}
        for history_rec in list(coproduct_history_layout.get("records") or []):
            history_quarter_end = history_rec.get("quarter_end")
            if isinstance(history_quarter_end, date):
                history_record_by_quarter[history_quarter_end] = dict(history_rec)
        available_credit_rows: List[Tuple[date, int]] = []
        if basis_proxy_ws is not None and history_first_row > 0 and history_last_row >= history_first_row and history_quarter_col > 0 and history_credit_per_gal_col > 0:
            for history_row in range(history_first_row, history_last_row + 1):
                quarter_label_txt = str(basis_proxy_ws.cell(row=history_row, column=history_quarter_col).value or "").strip()
                quarter_end = _parse_quarter_label_text(quarter_label_txt)
                credit_value = pd.to_numeric(basis_proxy_ws.cell(row=history_row, column=history_credit_per_gal_col).value, errors="coerce")
                if isinstance(quarter_end, date) and pd.notna(credit_value):
                    available_credit_rows.append((quarter_end, history_row))
        chart_credit_rows: List[Dict[str, Any]] = []
        if available_credit_rows:
            chart_credit_rows_by_end: Dict[date, Dict[str, Any]] = {}
            historical_credit_quarter_ends: Set[date] = set()
            frame_credit_col = int(frame_value_col_map.get("coproduct_credit_per_gal") or 0)
            frame_credit_usd_m_col = int(frame_value_col_map.get("coproduct_credit_usd_m") or 0)
            for quarter_end, history_row in available_credit_rows:
                quarter_label_txt = str(basis_proxy_ws.cell(row=history_row, column=history_quarter_col).value or "").strip()
                history_credit_formula = f"=Basis_Proxy_Sandbox!${get_column_letter(history_credit_per_gal_col)}${history_row}"
                history_credit_num = pd.to_numeric(
                    basis_proxy_ws.cell(row=history_row, column=history_credit_per_gal_col).value,
                    errors="coerce",
                )
                history_credit_usd_m_num = pd.to_numeric(
                    basis_proxy_ws.cell(row=history_row, column=history_credit_usd_m_col).value,
                    errors="coerce",
                ) if history_credit_usd_m_col > 0 else pd.NA
                history_coverage_num = pd.to_numeric(
                    basis_proxy_ws.cell(row=history_row, column=history_coverage_col).value,
                    errors="coerce",
                ) if history_coverage_col > 0 else pd.NA
                history_source_mode_txt = (
                    str(basis_proxy_ws.cell(row=history_row, column=history_source_mode_col).value or "").strip()
                    if history_source_mode_col > 0
                    else ""
                )
                history_credit_usd_m_formula = (
                    f"=Basis_Proxy_Sandbox!${get_column_letter(history_credit_usd_m_col)}${history_row}"
                    if history_credit_usd_m_col > 0
                    else ""
                )
                history_credit_usd_m_direct_write = False
                if pd.isna(history_credit_usd_m_num):
                    history_frame_key = str(frame_key_by_history_quarter.get(quarter_end) or "").strip()
                    history_frame_row = int(frame_rows_by_key.get(history_frame_key) or 0)
                    frame_credit_usd_m_num = (
                        pd.to_numeric(
                            basis_proxy_ws.cell(row=history_frame_row, column=frame_credit_usd_m_col).value,
                            errors="coerce",
                        )
                        if history_frame_row > 0 and frame_credit_usd_m_col > 0
                        else pd.NA
                    )
                    if pd.notna(frame_credit_usd_m_num):
                        history_credit_usd_m_num = frame_credit_usd_m_num
                        history_credit_usd_m_formula = ""
                        history_credit_usd_m_direct_write = True
                if not quarter_label_txt or pd.isna(history_credit_num):
                    continue
                chart_credit_rows_by_end[quarter_end] = {
                    "quarter_end": quarter_end,
                    "quarter_label": quarter_label_txt,
                    "quarter_formula": f"=Basis_Proxy_Sandbox!${get_column_letter(history_quarter_col)}${history_row}",
                    "formula": history_credit_formula,
                    "value": float(history_credit_num),
                    "usd_m_formula": history_credit_usd_m_formula,
                    "usd_m_value": (None if pd.isna(history_credit_usd_m_num) else float(history_credit_usd_m_num)),
                    "usd_m_direct_write": history_credit_usd_m_direct_write,
                    "coverage_formula": (
                        f"=Basis_Proxy_Sandbox!${get_column_letter(history_coverage_col)}${history_row}"
                        if history_coverage_col > 0
                        else ""
                    ),
                    "coverage_value": (None if pd.isna(history_coverage_num) else float(history_coverage_num)),
                    "source_mode_formula": (
                        f"=Basis_Proxy_Sandbox!${get_column_letter(history_source_mode_col)}${history_row}"
                        if history_source_mode_col > 0
                        else ""
                    ),
                    "source_mode_value": history_source_mode_txt,
                    "_preview_priority": -1,
                }
                if isinstance(prior_market_display_quarter, date) and quarter_end < prior_market_display_quarter:
                    historical_credit_quarter_ends.add(quarter_end)
            preview_priority = {
                "prior_quarter": 0,
                "quarter_open": 1,
                "current_qtd": 2,
                "next_quarter_thesis": 3,
            }
            ethanol_yield_num = pd.to_numeric(ws.cell(row=ethanol_yield_row, column=2).value, errors="coerce") if ethanol_yield_row > 0 else pd.NA
            corn_oil_yield_overlay_num = pd.to_numeric(
                (_overlay_coefficient_detail("renewable_corn_oil_yield") or {}).get("value"),
                errors="coerce",
            )
            for frame_key, target_quarter_end in (
                ("prior_quarter", prior_market_display_quarter),
                ("quarter_open", quarter_open_display_quarter),
                ("current_qtd", current_market_display_quarter),
                ("next_quarter_thesis", next_thesis_quarter_end),
            ):
                if not isinstance(target_quarter_end, date):
                    continue
                if target_quarter_end in historical_credit_quarter_ends:
                    continue
                frame_row_num = int(frame_rows_by_key.get(frame_key) or 0)
                if frame_row_num <= 0 or frame_credit_col <= 0:
                    continue
                frame_credit_num = pd.to_numeric(
                    basis_proxy_ws.cell(row=frame_row_num, column=frame_credit_col).value,
                    errors="coerce",
                )
                if pd.isna(frame_credit_num):
                    continue
                existing_rec = chart_credit_rows_by_end.get(target_quarter_end)
                if existing_rec is not None and str(existing_rec.get("quarter_formula") or "").strip():
                    continue
                if existing_rec is None:
                    existing_rec = {
                        "quarter_end": target_quarter_end,
                        "quarter_label": _quarter_label_short(target_quarter_end),
                        "quarter_formula": "",
                        "formula": "",
                        "value": None,
                        "usd_m_formula": "",
                        "coverage_formula": "",
                        "source_mode_formula": "",
                        "_preview_priority": -1,
                    }
                    chart_credit_rows_by_end[target_quarter_end] = existing_rec
                if int(existing_rec.get("_preview_priority") or -1) > int(preview_priority.get(frame_key, 0)):
                    continue
                frame_build_up_col = int(build_up_frame_col_by_key.get(frame_key) or 0)
                corn_oil_contribution_ref = (
                    f"Basis_Proxy_Sandbox!${get_column_letter(frame_build_up_col)}${build_up_corn_oil_contribution_row}"
                    if frame_build_up_col > 0 and build_up_corn_oil_contribution_row > 0
                    else ""
                )
                corn_oil_per_gal_formula = (
                    f'=IF(AND(ISNUMBER({corn_oil_contribution_ref}),ISNUMBER({ethanol_yield_ref}),ABS({ethanol_yield_ref})>1E-9),{corn_oil_contribution_ref}/{ethanol_yield_ref},"")'
                    if corn_oil_contribution_ref and ethanol_yield_ref
                    else ""
                )
                frame_total_credit_per_gal_ref = (
                    f"Basis_Proxy_Sandbox!${get_column_letter(frame_credit_col)}${frame_row_num}"
                    if frame_credit_col > 0
                    else ""
                )
                frame_total_credit_usd_m_ref = (
                    f"Basis_Proxy_Sandbox!${get_column_letter(frame_value_col_map.get('coproduct_credit_usd_m') or 0)}${frame_row_num}"
                    if int(frame_value_col_map.get("coproduct_credit_usd_m") or 0) > 0
                    else ""
                )
                frame_total_credit_usd_m_num = pd.to_numeric(
                    basis_proxy_ws.cell(row=frame_row_num, column=int(frame_value_col_map.get("coproduct_credit_usd_m") or 0)).value,
                    errors="coerce",
                ) if int(frame_value_col_map.get("coproduct_credit_usd_m") or 0) > 0 else pd.NA
                frame_coverage_num = pd.to_numeric(
                    basis_proxy_ws.cell(row=frame_row_num, column=frame_coverage_col).value,
                    errors="coerce",
                ) if frame_coverage_col > 0 else pd.NA
                frame_source_mode_txt = (
                    str(basis_proxy_ws.cell(row=frame_row_num, column=frame_source_mode_col).value or "").strip()
                    if frame_source_mode_col > 0
                    else ""
                )
                corn_oil_price_num = pd.to_numeric(
                    basis_proxy_ws.cell(row=frame_row_num, column=int(frame_value_col_map.get("corn_oil_price") or 0)).value,
                    errors="coerce",
                ) if int(frame_value_col_map.get("corn_oil_price") or 0) > 0 else pd.NA
                corn_oil_contribution_num = pd.to_numeric(
                    basis_proxy_ws.cell(row=build_up_corn_oil_contribution_row, column=frame_build_up_col).value,
                    errors="coerce",
                ) if build_up_corn_oil_contribution_row > 0 and frame_build_up_col > 0 else pd.NA
                corn_oil_per_gal_num = (
                    float(corn_oil_contribution_num) / float(ethanol_yield_num)
                    if pd.notna(corn_oil_contribution_num) and pd.notna(ethanol_yield_num) and abs(float(ethanol_yield_num)) > 1e-9
                    else None
                )
                if (
                    corn_oil_per_gal_num is None
                    and pd.notna(corn_oil_price_num)
                    and pd.notna(corn_oil_yield_overlay_num)
                    and pd.notna(ethanol_yield_num)
                    and abs(float(ethanol_yield_num)) > 1e-9
                ):
                    corn_oil_per_gal_num = (
                        float(corn_oil_price_num) * float(corn_oil_yield_overlay_num) / float(ethanol_yield_num)
                    )
                corn_oil_usd_m_proxy_num = (
                    float(corn_oil_per_gal_num) * (float(frame_total_credit_usd_m_num) / float(frame_credit_num))
                    if corn_oil_per_gal_num is not None
                    and pd.notna(frame_total_credit_usd_m_num)
                    and pd.notna(frame_credit_num)
                    and abs(float(frame_credit_num)) > 1e-9
                    else None
                )
                corn_oil_usd_m_proxy_formula = (
                    f'=IF(AND(ISNUMBER({corn_oil_contribution_ref}),ISNUMBER({ethanol_yield_ref}),ABS({ethanol_yield_ref})>1E-9,ISNUMBER({frame_total_credit_per_gal_ref}),ABS({frame_total_credit_per_gal_ref})>1E-9,ISNUMBER({frame_total_credit_usd_m_ref})),({corn_oil_contribution_ref}/{ethanol_yield_ref})*({frame_total_credit_usd_m_ref}/{frame_total_credit_per_gal_ref}),"")'
                    if corn_oil_contribution_ref and ethanol_yield_ref and frame_total_credit_per_gal_ref and frame_total_credit_usd_m_ref
                    else ""
                )
                existing_rec["_preview_priority"] = int(preview_priority.get(frame_key, 0))
                existing_rec["quarter_label"] = _quarter_label_short(target_quarter_end)
                existing_rec["formula"] = f"=Basis_Proxy_Sandbox!${get_column_letter(frame_credit_col)}${frame_row_num}"
                existing_rec["value"] = float(frame_credit_num)
                existing_rec["usd_m_value"] = None if pd.isna(frame_total_credit_usd_m_num) else float(frame_total_credit_usd_m_num)
                existing_rec["usd_m_formula"] = (
                    f"=Basis_Proxy_Sandbox!${get_column_letter(frame_value_col_map.get('coproduct_credit_usd_m') or 0)}${frame_row_num}"
                    if int(frame_value_col_map.get("coproduct_credit_usd_m") or 0) > 0
                    else ""
                )
                existing_rec["coverage_value"] = None if pd.isna(frame_coverage_num) else float(frame_coverage_num)
                existing_rec["coverage_formula"] = (
                    f"=Basis_Proxy_Sandbox!${get_column_letter(frame_coverage_col)}${frame_row_num}"
                    if frame_coverage_col > 0
                    else ""
                )
                existing_rec["source_mode_value"] = frame_source_mode_txt
                existing_rec["source_mode_formula"] = (
                    f"=Basis_Proxy_Sandbox!${get_column_letter(frame_source_mode_col)}${frame_row_num}"
                    if frame_source_mode_col > 0
                    else ""
                )
                existing_rec["frame_key"] = frame_key
                existing_rec["corn_oil_price_value"] = None if pd.isna(corn_oil_price_num) else float(corn_oil_price_num)
                existing_rec["corn_oil_price_formula"] = (
                    f"=Basis_Proxy_Sandbox!${get_column_letter(frame_value_col_map.get('corn_oil_price') or 0)}${frame_row_num}"
                    if int(frame_value_col_map.get("corn_oil_price") or 0) > 0
                    else ""
                )
                existing_rec["corn_oil_contribution_per_gal_value"] = corn_oil_per_gal_num
                existing_rec["corn_oil_contribution_per_gal_formula"] = corn_oil_per_gal_formula
                existing_rec["corn_oil_contribution_usd_m_proxy_value"] = corn_oil_usd_m_proxy_num
                existing_rec["corn_oil_contribution_usd_m_proxy_formula"] = corn_oil_usd_m_proxy_formula
            resolved_chart_credit_rows = [
                {
                    key: value
                    for key, value in rec.items()
                    if key != "_preview_priority"
                }
                for _, rec in sorted(chart_credit_rows_by_end.items(), key=lambda item: item[0])
                if isinstance(rec.get("quarter_end"), date)
                and str(rec.get("quarter_label") or "").strip()
                and pd.notna(pd.to_numeric(rec.get("value"), errors="coerce"))
                and str(rec.get("formula") or "").strip()
            ]
            if resolved_chart_credit_rows:
                chart_floor_quarter = date(2023, 3, 31)
                chart_credit_rows = [
                    rec
                    for rec in resolved_chart_credit_rows
                    if isinstance(rec.get("quarter_end"), date) and rec.get("quarter_end") >= chart_floor_quarter
                ]
                if len(chart_credit_rows) > overlay_quarter_chart_max_points:
                    chart_credit_rows = chart_credit_rows[-overlay_quarter_chart_max_points:]
        if basis_proxy_ws is not None and len(chart_credit_rows) >= 4:
            coproduct_chart_title_row = max(int(row_idx), visible_coproduct_rows["coproduct_credit_usd_m"] + 2)
            coproduct_chart_anchor_row = coproduct_chart_title_row + 1
            coproduct_chart_end_row = coproduct_chart_anchor_row + overlay_chart_row_span
            chart_start_col = 2  # B
            chart_end_col = 21  # U
            coproduct_helper_label_col = 49  # AW
            coproduct_helper_value_col = 50  # AX
            ws.column_dimensions[get_column_letter(coproduct_helper_label_col)].hidden = True
            ws.column_dimensions[get_column_letter(coproduct_helper_value_col)].hidden = True
            ws.merge_cells(
                start_row=coproduct_chart_title_row,
                start_column=chart_start_col,
                end_row=coproduct_chart_title_row,
                end_column=chart_end_col,
            )
            coproduct_chart_title = ws.cell(
                row=coproduct_chart_title_row,
                column=2,
                value="Approximate coproduct credit ($/gal, quarterly history)",
            )
            coproduct_chart_title.fill = title_fill
            coproduct_chart_title.font = title_font
            coproduct_chart_title.alignment = align_center
            coproduct_chart_title.border = thin_border
            for cc in range(chart_start_col, chart_end_col + 1):
                ws.cell(row=coproduct_chart_title_row, column=cc).fill = title_fill
                ws.cell(row=coproduct_chart_title_row, column=cc).font = title_font
                ws.cell(row=coproduct_chart_title_row, column=cc).alignment = align_center
                ws.cell(row=coproduct_chart_title_row, column=cc).border = thin_border
            ws.row_dimensions[coproduct_chart_title_row].height = 18.0

            coproduct_helper_start_row = coproduct_chart_title_row
            ws.cell(row=coproduct_helper_start_row, column=coproduct_helper_label_col, value="Quarter")
            ws.cell(
                row=coproduct_helper_start_row,
                column=coproduct_helper_value_col,
                value="Approximate coproduct credit ($/gal)",
            )
            coproduct_helper_last_row = coproduct_helper_start_row
            coproduct_chart_y_values: List[float] = []
            for helper_row, rec in enumerate(chart_credit_rows, start=coproduct_helper_start_row + 1):
                quarter_label_txt = str(rec.get("quarter_label") or "").strip()
                per_gal_num = pd.to_numeric(rec.get("value"), errors="coerce")
                helper_formula = str(rec.get("formula") or "").strip()
                if not quarter_label_txt or pd.isna(per_gal_num):
                    continue
                ws.cell(row=helper_row, column=coproduct_helper_label_col, value=quarter_label_txt)
                helper_value_cell = ws.cell(
                    row=helper_row,
                    column=coproduct_helper_value_col,
                    value=helper_formula or None,
                )
                helper_value_cell.number_format = "#,##0.000"
                coproduct_chart_y_values.append(float(per_gal_num))
                coproduct_helper_last_row = helper_row

            coproduct_chart = LineChart()
            coproduct_chart.title = None
            coproduct_chart.height = overlay_chart_height
            coproduct_chart.width = overlay_chart_width
            coproduct_chart.x_axis.title = None
            coproduct_chart.x_axis.tickLblPos = "low"
            coproduct_chart.x_axis.majorTickMark = "out"
            coproduct_chart.x_axis.minorTickMark = "none"
            coproduct_chart.y_axis.title = None
            coproduct_chart.y_axis.axPos = "l"
            coproduct_chart.y_axis.delete = False
            coproduct_chart.y_axis.number_format = "$0.000"
            coproduct_chart.y_axis.tickLblPos = "nextTo"
            coproduct_chart.y_axis.crosses = "min"
            coproduct_chart.y_axis.majorTickMark = "out"
            coproduct_chart.y_axis.minorTickMark = "none"
            coproduct_chart.visible_cells_only = False
            coproduct_chart.legend = None

            categories_ref = Reference(
                ws,
                min_col=coproduct_helper_label_col,
                min_row=coproduct_helper_start_row + 1,
                max_row=coproduct_helper_last_row,
            )
            values_ref = Reference(
                ws,
                min_col=coproduct_helper_value_col,
                min_row=coproduct_helper_start_row,
                max_row=coproduct_helper_last_row,
            )
            coproduct_chart.add_data(values_ref, titles_from_data=True, from_rows=False)
            _apply_chart_text_categories(
                coproduct_chart,
                sheet_name=ws.title,
                col_idx=coproduct_helper_label_col,
                start_row=coproduct_helper_start_row + 1,
                end_row=coproduct_helper_last_row,
            )
            try:
                coproduct_chart.series[0].graphicalProperties.line.solidFill = "2F80ED"
                coproduct_chart.series[0].graphicalProperties.line.width = 19050
                coproduct_chart.series[0].marker.symbol = "circle"
                coproduct_chart.series[0].marker.size = 7
                coproduct_chart.series[0].dLbls = DataLabelList(
                    showLegendKey=False,
                    showVal=True,
                    showSerName=False,
                    showCatName=False,
                    showLeaderLines=False,
                    dLblPos="b",
                    numFmt="#,##0.000",
                )
            except Exception:
                pass
            if coproduct_chart_y_values:
                coproduct_y_min = min(coproduct_chart_y_values)
                coproduct_y_max = max(coproduct_chart_y_values)
                coproduct_span = max(float(coproduct_y_max) - float(coproduct_y_min), 0.10)
                coproduct_pad = max(coproduct_span * 0.08, 0.03)
                try:
                    coproduct_chart.y_axis.scaling.min = float(coproduct_y_min) - coproduct_pad
                    coproduct_chart.y_axis.scaling.max = float(coproduct_y_max) + coproduct_pad
                except Exception:
                    pass
            coproduct_chart.anchor = TwoCellAnchor(
                _from=AnchorMarker(col=1, row=coproduct_chart_anchor_row - 1),
                to=AnchorMarker(col=chart_end_col, row=coproduct_chart_end_row),
            )
            ws.add_chart(coproduct_chart)
            mini_history_title_row = coproduct_chart_end_row + 2
            mini_history_header_row = mini_history_title_row + 1
            mini_history_first_data_row = mini_history_header_row + 1
            mini_history_col_spans = [
                (2, 3, "Quarter"),
                (4, 5, "$/gal"),
                (6, 7, "$m"),
                (8, 9, "Coverage"),
                (10, 11, "Source mode"),
            ]
            corn_oil_history_title_col = 13  # M
            corn_oil_history_title_end_col = 20  # T
            corn_oil_history_col_spans = [
                (13, 14, "Quarter"),
                (15, 16, "$/lb"),
                (17, 18, "$/gal"),
                (19, 20, "$m proxy"),
            ]
            mini_history_rows = list(reversed(chart_credit_rows))
            ws.merge_cells(start_row=mini_history_title_row, start_column=2, end_row=mini_history_title_row, end_column=11)
            mini_history_title = ws.cell(row=mini_history_title_row, column=2, value="Coproduct credit")
            mini_history_title.fill = copy(section_fill)
            mini_history_title.font = copy(bold_font)
            mini_history_title.alignment = Alignment(horizontal="center", vertical="center")
            mini_history_title.border = copy(thin_border)
            for cc in range(2, 12):
                ws.cell(row=mini_history_title_row, column=cc).fill = copy(section_fill)
                ws.cell(row=mini_history_title_row, column=cc).font = copy(bold_font)
                ws.cell(row=mini_history_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=mini_history_title_row, column=cc).border = copy(thin_border)
            ws.merge_cells(
                start_row=mini_history_title_row,
                start_column=corn_oil_history_title_col,
                end_row=mini_history_title_row,
                end_column=corn_oil_history_title_end_col,
            )
            corn_oil_history_title = ws.cell(row=mini_history_title_row, column=corn_oil_history_title_col, value="Corn oil prices")
            corn_oil_history_title.fill = copy(section_fill)
            corn_oil_history_title.font = copy(bold_font)
            corn_oil_history_title.alignment = Alignment(horizontal="center", vertical="center")
            corn_oil_history_title.border = copy(thin_border)
            for cc in range(corn_oil_history_title_col, corn_oil_history_title_end_col + 1):
                ws.cell(row=mini_history_title_row, column=cc).fill = copy(section_fill)
                ws.cell(row=mini_history_title_row, column=cc).font = copy(bold_font)
                ws.cell(row=mini_history_title_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=mini_history_title_row, column=cc).border = copy(thin_border)
            ws.row_dimensions[mini_history_title_row].height = 20.0

            for start_col, end_col, header_txt in mini_history_col_spans:
                if end_col > start_col:
                    ws.merge_cells(start_row=mini_history_header_row, start_column=start_col, end_row=mini_history_header_row, end_column=end_col)
                for cc in range(start_col, end_col + 1):
                    ws.cell(row=mini_history_header_row, column=cc).fill = copy(header_fill)
                    ws.cell(row=mini_history_header_row, column=cc).font = copy(bold_font)
                    ws.cell(row=mini_history_header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws.cell(row=mini_history_header_row, column=cc).border = copy(thin_border)
                ws.cell(row=mini_history_header_row, column=start_col, value=header_txt)
            for start_col, end_col, header_txt in corn_oil_history_col_spans:
                if end_col > start_col:
                    ws.merge_cells(start_row=mini_history_header_row, start_column=start_col, end_row=mini_history_header_row, end_column=end_col)
                for cc in range(start_col, end_col + 1):
                    ws.cell(row=mini_history_header_row, column=cc).fill = copy(header_fill)
                    ws.cell(row=mini_history_header_row, column=cc).font = copy(bold_font)
                    ws.cell(row=mini_history_header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws.cell(row=mini_history_header_row, column=cc).border = copy(thin_border)
                ws.cell(row=mini_history_header_row, column=start_col, value=header_txt)
            ws.row_dimensions[mini_history_header_row].height = 26.0

            latest_actual_coproduct_quarter = max(
                [
                    qd
                    for qd in history_record_by_quarter.keys()
                    if isinstance(qd, date)
                ],
                default=None,
            )
            for row_offset, rec in enumerate(mini_history_rows):
                target_row = mini_history_first_data_row + row_offset
                row_fill = copy(zebra_fill_light if ((target_row - mini_history_header_row) % 2) else zebra_fill_dark)
                for start_col, end_col, _header_txt in mini_history_col_spans:
                    if end_col > start_col:
                        ws.merge_cells(start_row=target_row, start_column=start_col, end_row=target_row, end_column=end_col)
                    for cc in range(start_col, end_col + 1):
                        ws.cell(row=target_row, column=cc).fill = copy(row_fill)
                        ws.cell(row=target_row, column=cc).font = copy(body_font)
                        ws.cell(row=target_row, column=cc).border = copy(thin_border)
                        ws.cell(row=target_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=start_col == 10)
                for start_col, end_col, _header_txt in corn_oil_history_col_spans:
                    if end_col > start_col:
                        ws.merge_cells(start_row=target_row, start_column=start_col, end_row=target_row, end_column=end_col)
                    for cc in range(start_col, end_col + 1):
                        ws.cell(row=target_row, column=cc).fill = copy(row_fill)
                        ws.cell(row=target_row, column=cc).font = copy(body_font)
                        ws.cell(row=target_row, column=cc).border = copy(thin_border)
                        ws.cell(row=target_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")
                history_quarter_end = rec.get("quarter_end")
                quarter_formula = str(rec.get("quarter_formula") or "").strip()
                quarter_label_txt = str(rec.get("quarter_label") or _quarter_label_short(history_quarter_end) or "").strip()
                is_preview_row = bool(str(rec.get("frame_key") or "").strip())
                is_forward_outlook_row = bool(
                    is_preview_row
                    or row_offset < 2
                    or (
                        isinstance(history_quarter_end, date)
                        and isinstance(latest_actual_coproduct_quarter, date)
                        and history_quarter_end > latest_actual_coproduct_quarter
                    )
                )
                if is_forward_outlook_row:
                    outlook_cell = ws.cell(row=target_row, column=12, value="Outlook")
                    outlook_cell.fill = copy(row_fill)
                    outlook_cell.border = copy(thin_border)
                    outlook_cell.font = Font(name=getattr(body_font, "name", None) or "Calibri", size=13, bold=True, color="1F4E79")
                    outlook_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=target_row, column=2, value=quarter_formula or quarter_label_txt)
                credit_formula = str(rec.get("formula") or "").strip()
                credit_num = pd.to_numeric(rec.get("value"), errors="coerce")
                if is_preview_row and pd.notna(credit_num):
                    ws.cell(row=target_row, column=4, value=float(credit_num)).number_format = "#,##0.000"
                elif credit_formula:
                    ws.cell(row=target_row, column=4, value=credit_formula).number_format = "#,##0.000"
                usd_m_formula = str(rec.get("usd_m_formula") or "").strip()
                usd_m_num = pd.to_numeric(rec.get("usd_m_value"), errors="coerce")
                usd_m_direct_write = bool(rec.get("usd_m_direct_write"))
                if pd.notna(usd_m_num) and (is_preview_row or usd_m_direct_write or not usd_m_formula):
                    ws.cell(row=target_row, column=6, value=float(usd_m_num)).number_format = "#,##0.0"
                elif usd_m_formula:
                    ws.cell(row=target_row, column=6, value=usd_m_formula).number_format = "#,##0.0"
                coverage_formula = str(rec.get("coverage_formula") or "").strip()
                coverage_num = pd.to_numeric(rec.get("coverage_value"), errors="coerce")
                if is_preview_row and pd.notna(coverage_num):
                    coverage_cell = ws.cell(row=target_row, column=8, value=float(coverage_num))
                    coverage_cell.number_format = "0%"
                elif coverage_formula:
                    coverage_cell = ws.cell(row=target_row, column=8, value=coverage_formula)
                    coverage_cell.number_format = "0%"
                source_mode_formula = str(rec.get("source_mode_formula") or "").strip()
                source_mode_value = str(rec.get("source_mode_value") or "").strip()
                if is_preview_row and source_mode_value:
                    ws.cell(row=target_row, column=10, value=source_mode_value)
                elif source_mode_formula:
                    ws.cell(row=target_row, column=10, value=source_mode_formula)
                history_rec = dict(history_record_by_quarter.get(history_quarter_end) or {})
                ws.cell(row=target_row, column=13, value=quarter_formula or quarter_label_txt)
                corn_oil_price_formula = str(rec.get("corn_oil_price_formula") or "").strip()
                corn_oil_per_gal_formula = str(rec.get("corn_oil_contribution_per_gal_formula") or "").strip()
                corn_oil_usd_m_proxy_formula = str(rec.get("corn_oil_contribution_usd_m_proxy_formula") or "").strip()
                corn_oil_price_num = pd.to_numeric(rec.get("corn_oil_price_value"), errors="coerce")
                corn_oil_per_gal_num = pd.to_numeric(rec.get("corn_oil_contribution_per_gal_value"), errors="coerce")
                corn_oil_usd_m_proxy_num = pd.to_numeric(rec.get("corn_oil_contribution_usd_m_proxy_value"), errors="coerce")
                if is_preview_row and pd.notna(corn_oil_price_num):
                    ws.cell(row=target_row, column=15, value=float(corn_oil_price_num)).number_format = "#,##0.000"
                elif corn_oil_price_formula:
                    ws.cell(row=target_row, column=15, value=corn_oil_price_formula).number_format = "#,##0.000"
                else:
                    corn_oil_price_num = pd.to_numeric(history_rec.get("renewable_corn_oil_price"), errors="coerce")
                    if pd.notna(corn_oil_price_num):
                        ws.cell(row=target_row, column=15, value=float(corn_oil_price_num)).number_format = "#,##0.000"
                if is_preview_row and pd.notna(corn_oil_per_gal_num):
                    ws.cell(row=target_row, column=17, value=float(corn_oil_per_gal_num)).number_format = "#,##0.000"
                elif corn_oil_per_gal_formula:
                    ws.cell(row=target_row, column=17, value=corn_oil_per_gal_formula).number_format = "#,##0.000"
                else:
                    corn_oil_contribution_per_gal_num = pd.to_numeric(
                        history_rec.get("renewable_corn_oil_contribution_per_gal"),
                        errors="coerce",
                    )
                    if pd.isna(corn_oil_contribution_per_gal_num):
                        corn_oil_contribution_per_bushel_num = pd.to_numeric(
                            history_rec.get("renewable_corn_oil_contribution_per_bushel"),
                            errors="coerce",
                        )
                        if (
                            pd.notna(corn_oil_contribution_per_bushel_num)
                            and pd.notna(ethanol_yield_num)
                            and abs(float(ethanol_yield_num)) > 1e-9
                        ):
                            corn_oil_contribution_per_gal_num = (
                                float(corn_oil_contribution_per_bushel_num) / float(ethanol_yield_num)
                            )
                    if pd.isna(corn_oil_contribution_per_gal_num):
                        history_corn_oil_price_num = pd.to_numeric(
                            history_rec.get("renewable_corn_oil_price"),
                            errors="coerce",
                        )
                        if (
                            pd.notna(history_corn_oil_price_num)
                            and pd.notna(corn_oil_yield_overlay_num)
                            and pd.notna(ethanol_yield_num)
                            and abs(float(ethanol_yield_num)) > 1e-9
                        ):
                            corn_oil_contribution_per_gal_num = (
                                float(history_corn_oil_price_num) * float(corn_oil_yield_overlay_num) / float(ethanol_yield_num)
                            )
                    if pd.notna(corn_oil_contribution_per_gal_num):
                        ws.cell(row=target_row, column=17, value=float(corn_oil_contribution_per_gal_num)).number_format = "#,##0.000"
                if is_preview_row and pd.notna(corn_oil_usd_m_proxy_num):
                    ws.cell(row=target_row, column=19, value=float(corn_oil_usd_m_proxy_num)).number_format = "#,##0.0"
                elif corn_oil_usd_m_proxy_formula:
                    ws.cell(row=target_row, column=19, value=corn_oil_usd_m_proxy_formula).number_format = "#,##0.0"
                else:
                    corn_oil_contribution_usd_m_proxy_num = pd.to_numeric(
                        history_rec.get("renewable_corn_oil_contribution_usd_m_proxy"),
                        errors="coerce",
                    )
                    if pd.isna(corn_oil_contribution_usd_m_proxy_num):
                        corn_oil_contribution_per_gal_num = pd.to_numeric(
                            ws.cell(row=target_row, column=17).value,
                            errors="coerce",
                        )
                        history_credit_per_gal_num = pd.to_numeric(
                            history_rec.get("approximate_coproduct_credit_per_gal"),
                            errors="coerce",
                        )
                        if pd.isna(history_credit_per_gal_num):
                            history_credit_per_gal_num = pd.to_numeric(rec.get("value"), errors="coerce")
                        history_credit_usd_m_num = pd.to_numeric(
                            history_rec.get("approximate_coproduct_credit_usd_m"),
                            errors="coerce",
                        )
                        if pd.isna(history_credit_usd_m_num):
                            history_credit_usd_m_num = pd.to_numeric(rec.get("usd_m_value"), errors="coerce")
                        if (
                            pd.notna(corn_oil_contribution_per_gal_num)
                            and pd.notna(history_credit_per_gal_num)
                            and abs(float(history_credit_per_gal_num)) > 1e-9
                            and pd.notna(history_credit_usd_m_num)
                        ):
                            corn_oil_contribution_usd_m_proxy_num = (
                                float(corn_oil_contribution_per_gal_num)
                                * (float(history_credit_usd_m_num) / float(history_credit_per_gal_num))
                            )
                    if (
                        pd.isna(corn_oil_contribution_usd_m_proxy_num)
                        and isinstance(history_quarter_end, date)
                        and basis_proxy_ws is not None
                    ):
                        history_frame_key = str(frame_key_by_history_quarter.get(history_quarter_end) or "").strip()
                        history_frame_row = int(frame_rows_by_key.get(history_frame_key) or 0)
                        history_frame_credit_col = int(frame_value_col_map.get("coproduct_credit_per_gal") or 0)
                        history_frame_credit_usd_m_col = int(frame_value_col_map.get("coproduct_credit_usd_m") or 0)
                        history_frame_corn_oil_price_col = int(frame_value_col_map.get("corn_oil_price") or 0)
                        history_frame_credit_num = (
                            pd.to_numeric(
                                basis_proxy_ws.cell(row=history_frame_row, column=history_frame_credit_col).value,
                                errors="coerce",
                            )
                            if history_frame_row > 0 and history_frame_credit_col > 0
                            else pd.NA
                        )
                        history_frame_credit_usd_m_num = (
                            pd.to_numeric(
                                basis_proxy_ws.cell(row=history_frame_row, column=history_frame_credit_usd_m_col).value,
                                errors="coerce",
                            )
                            if history_frame_row > 0 and history_frame_credit_usd_m_col > 0
                            else pd.NA
                        )
                        history_frame_corn_oil_price_num = (
                            pd.to_numeric(
                                basis_proxy_ws.cell(row=history_frame_row, column=history_frame_corn_oil_price_col).value,
                                errors="coerce",
                            )
                            if history_frame_row > 0 and history_frame_corn_oil_price_col > 0
                            else pd.NA
                        )
                        if (
                            pd.isna(corn_oil_contribution_per_gal_num)
                            and pd.notna(history_frame_corn_oil_price_num)
                            and pd.notna(corn_oil_yield_overlay_num)
                            and pd.notna(ethanol_yield_num)
                            and abs(float(ethanol_yield_num)) > 1e-9
                        ):
                            corn_oil_contribution_per_gal_num = (
                                float(history_frame_corn_oil_price_num) * float(corn_oil_yield_overlay_num) / float(ethanol_yield_num)
                            )
                            ws.cell(row=target_row, column=17, value=float(corn_oil_contribution_per_gal_num)).number_format = "#,##0.000"
                        if (
                            pd.notna(corn_oil_contribution_per_gal_num)
                            and pd.notna(history_frame_credit_num)
                            and abs(float(history_frame_credit_num)) > 1e-9
                            and pd.notna(history_frame_credit_usd_m_num)
                        ):
                            corn_oil_contribution_usd_m_proxy_num = (
                                float(corn_oil_contribution_per_gal_num)
                                * (float(history_frame_credit_usd_m_num) / float(history_frame_credit_num))
                            )
                    if pd.notna(corn_oil_contribution_usd_m_proxy_num):
                        ws.cell(row=target_row, column=19, value=float(corn_oil_contribution_usd_m_proxy_num)).number_format = "#,##0.0"
                ws.row_dimensions[target_row].height = 22.0
            coverage_note_row = mini_history_first_data_row + max(len(mini_history_rows), 1) + 1
            ws.merge_cells(start_row=coverage_note_row, start_column=2, end_row=coverage_note_row, end_column=11)
            coverage_note_cell = ws.cell(
                row=coverage_note_row,
                column=2,
                value="Coverage reflects covered active-capacity footprint; values are covered-footprint weighted averages.",
            )
            coverage_note_cell.fill = copy(intro_fill)
            coverage_note_cell.font = copy(body_font)
            coverage_note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            coverage_note_cell.border = copy(thin_border)
            for cc in range(2, 12):
                ws.cell(row=coverage_note_row, column=cc).fill = copy(intro_fill)
                ws.cell(row=coverage_note_row, column=cc).border = copy(thin_border)
            ws.row_dimensions[coverage_note_row].height = 22.0
            row_idx = max(row_idx, coverage_note_row + 1)
    _record_writer_substage("write_excel.drivers.render.economics_overlay.coproduct_block", overlay_coproduct_block_started)

    final_chart_count = len(getattr(ws, "_charts", []) or [])
    return GpreEconomicsOverlayCoproductResult(
        row_idx=row_idx,
        row_count=max(0, int(row_idx) - int(initial_row_idx)),
        chart_count=max(0, int(final_chart_count) - int(initial_chart_count)),
    )
