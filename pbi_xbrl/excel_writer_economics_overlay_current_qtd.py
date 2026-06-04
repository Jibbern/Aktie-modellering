"""GPRE Current QTD tracking section for the Economics_Overlay sheet."""
from __future__ import annotations

import time
from copy import copy
from dataclasses import dataclass
from datetime import date
from typing import Any, Callable, Mapping

import pandas as pd
from openpyxl.styles import Alignment, Border, PatternFill


@dataclass(frozen=True)
class GpreEconomicsOverlayCurrentQtdDeps:
    ws: Any
    row_idx: int
    is_gpre_profile: bool
    has_gpre_commercial_setup: bool
    gpre_current_qtd_trend_tracking: Mapping[str, Any]
    title_fill: Any
    title_font: Any
    header_fill: Any
    body_font: Any
    bold_font: Any
    thin_border: Any
    zebra_fill_light: Any
    zebra_fill_dark: Any
    intro_fill: Any
    align_center: Any
    parse_snapshot_date_like: Callable[[Any], Any]
    record_writer_substage: Callable[[str, float], None]


@dataclass(frozen=True)
class GpreEconomicsOverlayCurrentQtdResult:
    row_idx: int
    row_count: int
    coproduct_start_row: int


def write_gpre_economics_overlay_current_qtd_section(
    deps: GpreEconomicsOverlayCurrentQtdDeps,
) -> GpreEconomicsOverlayCurrentQtdResult:
    ws = deps.ws
    row_idx = int(deps.row_idx)
    gpre_overlay_coproduct_start_row = 176
    overlay_qtd_tracking_started = time.perf_counter()
    row_count = 0
    if (
        deps.is_gpre_profile
        and deps.has_gpre_commercial_setup
        and isinstance(deps.gpre_current_qtd_trend_tracking, dict)
        and deps.gpre_current_qtd_trend_tracking
    ):
        current_qtd_tracking = dict(deps.gpre_current_qtd_trend_tracking.get("current_snapshot") or {})
        reference_comparisons = {
            str(key): dict(value or {})
            for key, value in dict(deps.gpre_current_qtd_trend_tracking.get("reference_comparisons") or {}).items()
        }
        driver_rows = [
            dict(rec or {})
            for rec in list(deps.gpre_current_qtd_trend_tracking.get("driver_attribution_rows") or [])
            if isinstance(rec, dict)
        ]
        qtd_tracking_start_row = max(int(row_idx) + 1, 176)
        qtd_tracking_title_row = qtd_tracking_start_row
        qtd_tracking_today_row = qtd_tracking_title_row + 1
        qtd_tracking_quarter_open_row = qtd_tracking_today_row + 1
        qtd_tracking_compare_header_row = qtd_tracking_quarter_open_row + 1
        qtd_tracking_compare_subheader_row = qtd_tracking_compare_header_row + 1
        qtd_tracking_compare_body_row = qtd_tracking_compare_subheader_row + 1
        qtd_tracking_spacer_row = qtd_tracking_compare_body_row + 1
        qtd_tracking_driver_title_row = qtd_tracking_spacer_row + 1
        qtd_tracking_driver_header_row = qtd_tracking_driver_title_row + 1
        qtd_tracking_driver_first_row = qtd_tracking_driver_header_row + 1
        qtd_tracking_driver_last_row = qtd_tracking_driver_first_row + max(len(driver_rows), 1) - 1
        qtd_tracking_note_row = qtd_tracking_driver_last_row + 1

        ws.merge_cells(
            start_row=qtd_tracking_title_row,
            start_column=1,
            end_row=qtd_tracking_title_row,
            end_column=21,
        )
        qtd_tracking_title_cell = ws.cell(
            row=qtd_tracking_title_row,
            column=1,
            value="Current QTD trend tracking ($/gal, crush margin lens)",
        )
        qtd_tracking_title_cell.fill = deps.title_fill
        qtd_tracking_title_cell.font = deps.title_font
        qtd_tracking_title_cell.alignment = deps.align_center
        qtd_tracking_title_cell.border = deps.thin_border
        for cc in range(1, 22):
            ws.cell(row=qtd_tracking_title_row, column=cc).fill = deps.title_fill
            ws.cell(row=qtd_tracking_title_row, column=cc).font = deps.title_font
            ws.cell(row=qtd_tracking_title_row, column=cc).alignment = deps.align_center
            ws.cell(row=qtd_tracking_title_row, column=cc).border = deps.thin_border
        ws.row_dimensions[qtd_tracking_title_row].height = 18.0

        def _qtd_as_of_text(raw_date: Any) -> str:
            parsed = deps.parse_snapshot_date_like(raw_date)
            return f"As of {parsed.isoformat()}" if isinstance(parsed, date) else "—"

        def _write_merged_numeric_or_dash(target_row: int, start_col: int, end_col: int, raw_value: Any) -> None:
            ws.merge_cells(
                start_row=target_row,
                start_column=start_col,
                end_row=target_row,
                end_column=end_col,
            )
            value_num = pd.to_numeric(raw_value, errors="coerce")
            if pd.notna(value_num):
                ws.cell(row=target_row, column=start_col, value=float(value_num)).number_format = "#,##0.000"
            else:
                ws.cell(row=target_row, column=start_col, value="—")

        def _write_merged_text(target_row: int, start_col: int, end_col: int, text_value: str) -> None:
            ws.merge_cells(
                start_row=target_row,
                start_column=start_col,
                end_row=target_row,
                end_column=end_col,
            )
            ws.cell(row=target_row, column=start_col, value=text_value)

        status_rows = [
            (
                qtd_tracking_today_row,
                "Today",
                pd.to_numeric(current_qtd_tracking.get("current_qtd_official_simple_usd_per_gal"), errors="coerce"),
                _qtd_as_of_text(current_qtd_tracking.get("as_of_date")),
                copy(deps.zebra_fill_light),
            ),
            (
                qtd_tracking_quarter_open_row,
                "Quarter-open",
                pd.to_numeric((reference_comparisons.get("quarter_open") or {}).get("reference_value_usd_per_gal"), errors="coerce"),
                _qtd_as_of_text((reference_comparisons.get("quarter_open") or {}).get("reference_date")),
                copy(deps.zebra_fill_dark),
            ),
        ]
        for target_row, label_txt, level_val, as_of_txt, row_fill in status_rows:
            _write_merged_numeric_or_dash(target_row, 2, 3, level_val)
            _write_merged_text(target_row, 4, 21, as_of_txt)
            for cc in range(1, 22):
                ws.cell(row=target_row, column=cc).fill = copy(row_fill)
                ws.cell(row=target_row, column=cc).font = copy(deps.body_font)
                ws.cell(row=target_row, column=cc).border = copy(deps.thin_border)
                ws.cell(row=target_row, column=cc).alignment = (
                    Alignment(horizontal="left", vertical="center", wrap_text=True)
                    if cc in {1, 4}
                    else Alignment(horizontal="center", vertical="center")
                )
            ws.cell(row=target_row, column=1, value=label_txt)
            ws.cell(row=target_row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[target_row].height = max(float(ws.row_dimensions[target_row].height or 0.0), 20.0)

        ws.row_dimensions[qtd_tracking_spacer_row].height = 8.0
        ws.cell(row=qtd_tracking_spacer_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

        comparison_spans = [
            (2, 3, "QTD vs quarter-open"),
            (4, 5, "QTD vs 1 week ago"),
            (6, 7, "QTD vs 4 weeks ago"),
            (8, 9, "QTD vs 8 weeks ago"),
        ]
        comparison_keys = {
            2: "quarter_open",
            4: "1w",
            6: "4w",
            8: "8w",
        }
        for start_col, end_col, header_txt in comparison_spans:
            ws.merge_cells(
                start_row=qtd_tracking_compare_header_row,
                start_column=start_col,
                end_row=qtd_tracking_compare_header_row,
                end_column=end_col,
            )
            ws.merge_cells(
                start_row=qtd_tracking_compare_subheader_row,
                start_column=start_col,
                end_row=qtd_tracking_compare_subheader_row,
                end_column=end_col,
            )
            ws.merge_cells(
                start_row=qtd_tracking_compare_body_row,
                start_column=start_col,
                end_row=qtd_tracking_compare_body_row,
                end_column=end_col,
            )
            ws.cell(row=qtd_tracking_compare_header_row, column=start_col, value=header_txt)
            ref_date_txt = _qtd_as_of_text((reference_comparisons.get(comparison_keys[start_col]) or {}).get("reference_date"))
            ws.cell(row=qtd_tracking_compare_subheader_row, column=start_col, value=ref_date_txt)
            delta_num = pd.to_numeric((reference_comparisons.get(comparison_keys[start_col]) or {}).get("delta_usd_per_gal"), errors="coerce")
            if pd.notna(delta_num):
                ws.cell(row=qtd_tracking_compare_body_row, column=start_col, value=float(delta_num)).number_format = "#,##0.000"
            else:
                ws.cell(row=qtd_tracking_compare_body_row, column=start_col, value="—")

        for cc in range(1, 22):
            ws.cell(row=qtd_tracking_compare_header_row, column=cc).fill = copy(deps.header_fill)
            ws.cell(row=qtd_tracking_compare_header_row, column=cc).font = copy(deps.bold_font)
            ws.cell(row=qtd_tracking_compare_header_row, column=cc).border = copy(deps.thin_border)
            ws.cell(row=qtd_tracking_compare_header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.cell(row=qtd_tracking_compare_subheader_row, column=cc).fill = copy(deps.intro_fill)
            ws.cell(row=qtd_tracking_compare_subheader_row, column=cc).font = copy(deps.body_font)
            ws.cell(row=qtd_tracking_compare_subheader_row, column=cc).border = copy(deps.thin_border)
            ws.cell(row=qtd_tracking_compare_subheader_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.cell(row=qtd_tracking_compare_body_row, column=cc).fill = copy(deps.zebra_fill_light)
            ws.cell(row=qtd_tracking_compare_body_row, column=cc).font = copy(deps.body_font)
            ws.cell(row=qtd_tracking_compare_body_row, column=cc).border = copy(deps.thin_border)
            ws.cell(row=qtd_tracking_compare_body_row, column=cc).alignment = (
                Alignment(horizontal="left", vertical="center", wrap_text=True)
                if cc == 1
                else Alignment(horizontal="center", vertical="center")
            )
        ws.cell(row=qtd_tracking_compare_body_row, column=1, value="Approximate market crush")
        ws.row_dimensions[qtd_tracking_compare_header_row].height = 24.0
        ws.row_dimensions[qtd_tracking_compare_subheader_row].height = 20.0
        ws.row_dimensions[qtd_tracking_compare_body_row].height = 20.0

        for quiet_row in range(qtd_tracking_compare_body_row + 1, qtd_tracking_driver_title_row):
            ws.row_dimensions[quiet_row].height = 0.0
            ws.row_dimensions[quiet_row].hidden = True

        ws.merge_cells(
            start_row=qtd_tracking_driver_title_row,
            start_column=1,
            end_row=qtd_tracking_driver_title_row,
            end_column=21,
        )
        driver_title_cell = ws.cell(
            row=qtd_tracking_driver_title_row,
            column=1,
            value="Driver attribution of Current QTD move ($/gal)",
        )
        driver_title_cell.fill = copy(deps.header_fill)
        driver_title_cell.font = copy(deps.bold_font)
        driver_title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        driver_title_cell.border = copy(deps.thin_border)
        for cc in range(1, 22):
            ws.cell(row=qtd_tracking_driver_title_row, column=cc).fill = copy(deps.header_fill)
            ws.cell(row=qtd_tracking_driver_title_row, column=cc).font = copy(deps.bold_font)
            ws.cell(row=qtd_tracking_driver_title_row, column=cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row=qtd_tracking_driver_title_row, column=cc).border = copy(deps.thin_border)
        ws.row_dimensions[qtd_tracking_driver_title_row].height = 20.0

        driver_header_spans = [
            (1, 1, "Driver"),
            (2, 3, "QTD vs quarter-open"),
            (4, 5, "QTD vs 1 week ago"),
            (6, 7, "QTD vs 4 weeks ago"),
            (8, 9, "QTD vs 8 weeks ago"),
        ]
        for start_col, end_col, header_txt in driver_header_spans:
            if end_col > start_col:
                ws.merge_cells(
                    start_row=qtd_tracking_driver_header_row,
                    start_column=start_col,
                    end_row=qtd_tracking_driver_header_row,
                    end_column=end_col,
                )
            for cc in range(start_col, end_col + 1):
                ws.cell(row=qtd_tracking_driver_header_row, column=cc).fill = copy(deps.header_fill)
                ws.cell(row=qtd_tracking_driver_header_row, column=cc).font = copy(deps.bold_font)
                ws.cell(row=qtd_tracking_driver_header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.cell(row=qtd_tracking_driver_header_row, column=cc).border = copy(deps.thin_border)
            ws.cell(row=qtd_tracking_driver_header_row, column=start_col, value=header_txt)
        for cc in range(10, 22):
            ws.cell(row=qtd_tracking_driver_header_row, column=cc).fill = copy(deps.header_fill)
            ws.cell(row=qtd_tracking_driver_header_row, column=cc).border = copy(deps.thin_border)
        ws.row_dimensions[qtd_tracking_driver_header_row].height = 24.0

        driver_key_by_col = {
            2: "quarter_open",
            4: "1w",
            6: "4w",
            8: "8w",
        }
        for row_offset, rec in enumerate(driver_rows or [{"driver": "Gas"}]):
            target_row = qtd_tracking_driver_first_row + row_offset
            row_fill = copy(deps.zebra_fill_light if (row_offset % 2 == 0) else deps.zebra_fill_dark)
            ws.merge_cells(start_row=target_row, start_column=2, end_row=target_row, end_column=3)
            ws.merge_cells(start_row=target_row, start_column=4, end_row=target_row, end_column=5)
            ws.merge_cells(start_row=target_row, start_column=6, end_row=target_row, end_column=7)
            ws.merge_cells(start_row=target_row, start_column=8, end_row=target_row, end_column=9)
            for cc in range(1, 22):
                ws.cell(row=target_row, column=cc).fill = copy(row_fill)
                ws.cell(row=target_row, column=cc).font = copy(deps.body_font)
                ws.cell(row=target_row, column=cc).border = copy(deps.thin_border)
                ws.cell(row=target_row, column=cc).alignment = (
                    Alignment(horizontal="left", vertical="center", wrap_text=True)
                    if cc == 1
                    else Alignment(horizontal="center", vertical="center")
                )
            ws.cell(row=target_row, column=1, value=str(rec.get("driver") or ""))
            for col_num, ref_key in driver_key_by_col.items():
                delta_num = pd.to_numeric(rec.get(ref_key), errors="coerce")
                if pd.notna(delta_num):
                    ws.cell(row=target_row, column=col_num, value=float(delta_num)).number_format = "#,##0.000"
            if str(rec.get("driver") or "").strip().lower() == "corn basis":
                ws.merge_cells(start_row=target_row, start_column=10, end_row=target_row, end_column=21)
                basis_note_cell = ws.cell(
                    row=target_row,
                    column=10,
                    value="Lower (more negative) corn basis lowers delivered corn cost and supports crush margins.",
                )
                basis_note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                basis_note_cell.font = copy(deps.body_font)
                ws.row_dimensions[target_row].height = max(float(ws.row_dimensions[target_row].height or 20.0), 24.0)
            ws.row_dimensions[target_row].height = 20.0

        for cc in range(1, 22):
            ws.cell(row=qtd_tracking_note_row, column=cc).fill = PatternFill(fill_type=None)
            ws.cell(row=qtd_tracking_note_row, column=cc).border = Border()
        ws.row_dimensions[qtd_tracking_note_row].height = 15.0
        row_idx = max(row_idx, qtd_tracking_note_row + 1)
        gpre_overlay_coproduct_start_row = qtd_tracking_note_row + 2
        row_count = qtd_tracking_note_row - qtd_tracking_start_row + 1
    deps.record_writer_substage(
        "write_excel.drivers.render.economics_overlay.current_qtd_tracking",
        overlay_qtd_tracking_started,
    )
    return GpreEconomicsOverlayCurrentQtdResult(
        row_idx=row_idx,
        row_count=row_count,
        coproduct_start_row=gpre_overlay_coproduct_start_row,
    )
