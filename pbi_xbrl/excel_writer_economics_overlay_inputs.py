"""Economics_Overlay input row writers extracted from excel_writer_context."""

from __future__ import annotations

import time
from copy import copy
from dataclasses import dataclass
from datetime import date
from typing import Any, Callable, Dict, Mapping, Optional, Sequence

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _formula_if_ready(refs: Sequence[str], expr: str) -> str:
    use_refs = [ref for ref in refs if ref]
    if not use_refs:
        return '= ""'.replace(' ', '')
    checks = ",".join([f"ISNUMBER({ref})" for ref in use_refs])
    return f'=IFERROR(IF(AND({checks}),{expr},""),"")'


def _sheet_ref(ref_in: str, source_sheet_name: str = "") -> str:
    ref_txt = str(ref_in or "").strip()
    if not ref_txt:
        return ""
    return f"{source_sheet_name}!{ref_txt}" if source_sheet_name else ref_txt


@dataclass(frozen=True)
class GpreEconomicsOverlayBuildUpDeps:
    coeff_rows: Mapping[str, int]
    market_rows: Mapping[str, int]
    section_fill: Any
    bold_font: Any
    align_center: Any
    thin_border: Any
    intro_fill: Any
    body_font: Any
    align_left_top_wrap: Any
    header_fill: Any
    align_center_wrap: Any
    zebra_fill_light: Any
    zebra_fill_dark: Any
    align_left_center_wrap: Any
    prior_market_display_quarter_txt: str
    quarter_open_overlay_header_txt: str
    current_qtd_market_snapshot: Mapping[str, Any]
    next_thesis_quarter_txt: str
    prior_q_market_snapshot: Mapping[str, Any]
    quarter_open_market_snapshot: Mapping[str, Any]
    next_quarter_thesis_snapshot: Mapping[str, Any]
    overlay_as_of_header_text: Callable[..., str]
    official_corn_basis_snapshot_display: Callable[..., str]
    official_corn_basis_selection_rule_display: Callable[..., str]


@dataclass(frozen=True)
class GpreEconomicsOverlayInputRowsDeps:
    ws: Any
    row_idx: int
    is_gpre_profile: bool
    gpre_commercial_setup_rows: Sequence[Mapping[str, Any]]
    overlay_gpre_end_col: int
    overlay_display_quarters: Sequence[Any]
    coefficient_templates: Sequence[Any]
    gpre_official_market_rows: Sequence[Mapping[str, Any]]
    gpre_basis_quarter_map: Mapping[Any, Mapping[str, Any]]
    gpre_official_market_summary: str
    gpre_official_weighting_method: str
    gpre_official_ethanol_method: str
    gpre_official_basis_method: str
    gpre_official_gas_method: str
    gpre_official_fallback_policy: str
    hidden_overlay_coefficient_keys: set[str]
    hidden_overlay_market_input_keys: set[str]
    prior_market_display_quarter_txt: str
    quarter_open_overlay_header_txt: str
    current_qtd_market_snapshot: Mapping[str, Any]
    next_thesis_quarter_txt: str
    prior_q_market_snapshot: Mapping[str, Any]
    quarter_open_market_snapshot: Mapping[str, Any]
    next_quarter_thesis_snapshot: Mapping[str, Any]
    current_overlay_model_key: str
    overlay_section_row_height: float
    overlay_intro_row_height: float
    overlay_header_row_height: float
    overlay_support_row_height: float
    thin_border: Any
    body_font: Any
    input_fill: Any
    input_font: Any
    section_fill: Any
    align_left_center_wrap: Any
    align_center: Any
    align_left_center: Any
    align_center_wrap: Any
    align_left_top_wrap: Any
    intro_fill: Any
    zebra_fill_light: Any
    zebra_fill_dark: Any
    bold_font: Any
    header_fill: Any
    font_size: float
    dark_text_color: str
    write_section_bar: Callable[..., int]
    write_overlay_intro: Callable[..., int]
    write_header_row: Callable[..., int]
    center_header_span: Callable[..., None]
    write_overlay_subheader_row: Callable[..., int]
    overlay_coefficient_detail: Callable[..., Mapping[str, Any]]
    overlay_coefficient_basis_display: Callable[..., str]
    overlay_coefficient_source_display: Callable[..., str]
    add_comment: Callable[..., None]
    record_writer_substage: Callable[[str, float], None]
    market_input_intro_text: Callable[[], str]
    ordered_market_input_templates: Callable[[], Sequence[Any]]
    pick_market_reference: Callable[..., Mapping[str, Any]]
    market_source_note: Callable[..., str]
    driver_source_note: Callable[..., str]
    prior_market_override: Callable[..., Any]
    quarter_open_market_override: Callable[..., Any]
    current_market_override: Callable[..., Any]
    thesis_market_override: Callable[..., Any]
    market_input_source_text: Callable[..., str]
    overlay_as_of_header_text: Callable[..., str]
    overlay_preview_bundle_for_model: Callable[..., Mapping[str, Any]]
    snapshot_market_meta: Callable[..., Mapping[str, Any]]
    market_gpre_phase_preview_story: Callable[..., Mapping[str, Any]]


@dataclass(frozen=True)
class GpreEconomicsOverlayInputRowsResult:
    row_idx: int
    coeff_rows: Dict[str, int]
    market_rows: Dict[str, int]
    coeff_ref: Dict[str, str]
    prior_ref: Dict[str, str]
    quarter_open_ref: Dict[str, str]
    current_ref: Dict[str, str]
    thesis_ref: Dict[str, str]
    market_section_bar_row: Optional[int]
    market_inputs_started: float
    write_gpre_approx_market_crush_build_up_section: Callable[..., Dict[str, Any]]
    gpre_fitted_live_formula: Callable[..., Optional[str]]
    gpre_formula_note: Callable[..., str]
    gpre_model_live_formula: Callable[..., Optional[str]]
    gpre_model_formula_note: Callable[..., str]


def write_gpre_approx_market_crush_build_up_section(
    deps: GpreEconomicsOverlayBuildUpDeps,
    target_ws: Any,
    start_row_in: int,
    *,
    source_sheet_name: str,
) -> Dict[str, Any]:
    coeff_rows = deps.coeff_rows
    market_rows = deps.market_rows
    section_fill = deps.section_fill
    bold_font = deps.bold_font
    align_center = deps.align_center
    thin_border = deps.thin_border
    intro_fill = deps.intro_fill
    body_font = deps.body_font
    align_left_top_wrap = deps.align_left_top_wrap
    header_fill = deps.header_fill
    align_center_wrap = deps.align_center_wrap
    zebra_fill_light = deps.zebra_fill_light
    zebra_fill_dark = deps.zebra_fill_dark
    align_left_center_wrap = deps.align_left_center_wrap
    prior_market_display_quarter_txt = deps.prior_market_display_quarter_txt
    quarter_open_overlay_header_txt = deps.quarter_open_overlay_header_txt
    current_qtd_market_snapshot = deps.current_qtd_market_snapshot
    next_thesis_quarter_txt = deps.next_thesis_quarter_txt
    prior_q_market_snapshot = deps.prior_q_market_snapshot
    quarter_open_market_snapshot = deps.quarter_open_market_snapshot
    next_quarter_thesis_snapshot = deps.next_quarter_thesis_snapshot
    _overlay_as_of_header_text = deps.overlay_as_of_header_text
    _official_corn_basis_snapshot_display = deps.official_corn_basis_snapshot_display
    _official_corn_basis_selection_rule_display = deps.official_corn_basis_selection_rule_display
    if not (coeff_rows and market_rows):
        return {}
    build_ws = target_ws
    section_start_col = 2  # B
    section_end_col = 15  # O
    title_row = start_row_in
    note_row = title_row + 1
    header_row = note_row + 1
    subheader_row = header_row + 1
    econ_rows_local = {
        "ethanol_revenue": subheader_row + 1,
        "distillers_contribution": subheader_row + 2,
        "uhp_contribution": subheader_row + 3,
        "corn_oil_contribution": subheader_row + 4,
        "feedstock_cost": subheader_row + 5,
        "natural_gas_burden": subheader_row + 6,
        "coproduct_credit": subheader_row + 7,
        "process_margin": subheader_row + 8,
        "corn_basis_snapshot_date": subheader_row + 9,
        "corn_basis_selection_rule": subheader_row + 10,
    }
    econ_defs_local = [
        ("ethanol_revenue", "Ethanol revenue contribution", "$/bushel", "Yield * ethanol price"),
        ("distillers_contribution", "Distillers contribution", "$/bushel", "Yield * distillers price"),
        ("uhp_contribution", "Ultra-high protein contribution", "$/bushel", "Yield * Ultra-high protein price"),
        ("corn_oil_contribution", "Renewable corn oil contribution", "$/bushel", "Yield * renewable corn oil price"),
        ("feedstock_cost", "Feedstock cost", "$/bushel", "Corn price per bushel"),
        ("natural_gas_burden", "Natural gas burden", "$/bushel", "Natural gas usage * gas price"),
        ("coproduct_credit", "Approximate coproduct credit", "$/bushel", "Distillers + Ultra-high protein + corn oil contributions"),
        ("process_margin", "Approximate market crush", "$/gal", "Market crush estimate with natural gas cost and GPRE corn basis, weighted to active capacity, and converted to $/gal."),
        ("corn_basis_snapshot_date", "Official corn basis snapshot date", "date/text", "Retained GPRE corn-bid snapshot date used by the official corn-basis leg only; AMS fallback when no eligible retained snapshot exists."),
        ("corn_basis_selection_rule", "Official corn basis selection rule", "rule/text", "Frame-specific retained-snapshot selector used by the official corn-basis leg; AMS fallback appears only when no eligible retained snapshot exists."),
    ]
    source_coeff_ref = {k: _sheet_ref(f"$B${r}", source_sheet_name) for k, r in coeff_rows.items()}
    source_prior_ref = {k: _sheet_ref(f"$B${r}", source_sheet_name) for k, r in market_rows.items()}
    source_quarter_open_ref = {k: _sheet_ref(f"$D${r}", source_sheet_name) for k, r in market_rows.items()}
    source_current_ref = {k: _sheet_ref(f"$F${r}", source_sheet_name) for k, r in market_rows.items()}
    source_thesis_ref = {k: _sheet_ref(f"$H${r}", source_sheet_name) for k, r in market_rows.items()}
    frame_cols = {
        "prior_quarter": 3,
        "quarter_open": 5,
        "current_qtd": 7,
        "next_quarter_thesis": 9,
    }
    note_intro = "Official simple row build-up used by Approximate market crush on Economics_Overlay."
    build_ws.merge_cells(start_row=title_row, start_column=section_start_col, end_row=title_row, end_column=section_end_col)
    title_cell = build_ws.cell(row=title_row, column=section_start_col, value="Approximate market crush build-up ($/gal)")
    title_cell.fill = section_fill
    title_cell.font = bold_font
    title_cell.alignment = align_center
    title_cell.border = thin_border
    for cc in range(section_start_col, section_end_col + 1):
        build_ws.cell(row=title_row, column=cc).fill = section_fill
        build_ws.cell(row=title_row, column=cc).font = bold_font
        build_ws.cell(row=title_row, column=cc).alignment = align_center
        build_ws.cell(row=title_row, column=cc).border = thin_border
    build_ws.row_dimensions[title_row].height = 22.0

    build_ws.merge_cells(start_row=note_row, start_column=section_start_col, end_row=note_row, end_column=section_end_col)
    note_cell = build_ws.cell(row=note_row, column=section_start_col, value=note_intro)
    note_cell.fill = intro_fill
    note_cell.font = body_font
    note_cell.alignment = align_left_top_wrap
    note_cell.border = thin_border
    for cc in range(section_start_col, section_end_col + 1):
        build_ws.cell(row=note_row, column=cc).fill = intro_fill
        build_ws.cell(row=note_row, column=cc).border = thin_border
    build_ws.row_dimensions[note_row].height = 24.0

    header_spans = [
        (2, 2, "Line item"),
        (3, 4, "Prior quarter"),
        (5, 6, "Quarter-open outlook"),
        (7, 8, "Current QTD"),
        (9, 10, "Next quarter outlook"),
        (11, 11, "Unit"),
        (12, 15, "Note"),
    ]
    for start_col, end_col, hdr in header_spans:
        if end_col > start_col:
            build_ws.merge_cells(start_row=header_row, start_column=start_col, end_row=header_row, end_column=end_col)
        for cc in range(start_col, end_col + 1):
            build_ws.cell(row=header_row, column=cc).fill = header_fill
            build_ws.cell(row=header_row, column=cc).font = bold_font
            build_ws.cell(row=header_row, column=cc).border = thin_border
            build_ws.cell(row=header_row, column=cc).alignment = align_center_wrap
        build_ws.cell(row=header_row, column=start_col, value=hdr)
    build_ws.row_dimensions[header_row].height = 24.0

    subheader_values = {
        3: prior_market_display_quarter_txt,
        5: quarter_open_overlay_header_txt,
        7: _overlay_as_of_header_text(current_qtd_market_snapshot.get("process_as_of") if isinstance(current_qtd_market_snapshot, dict) else None),
        9: next_thesis_quarter_txt,
    }
    for start_col, end_col in ((3, 4), (5, 6), (7, 8), (9, 10), (12, 15)):
        if end_col > start_col:
            build_ws.merge_cells(start_row=subheader_row, start_column=start_col, end_row=subheader_row, end_column=end_col)
    for cc in range(section_start_col, section_end_col + 1):
        build_ws.cell(row=subheader_row, column=cc).fill = zebra_fill_light
        build_ws.cell(row=subheader_row, column=cc).font = body_font
        build_ws.cell(row=subheader_row, column=cc).border = thin_border
        build_ws.cell(row=subheader_row, column=cc).alignment = align_center_wrap if cc in {3, 5, 7, 9, 11} else align_left_center_wrap
    for col_idx, subheader_txt in subheader_values.items():
        build_ws.cell(row=subheader_row, column=col_idx, value=subheader_txt)
    build_ws.row_dimensions[subheader_row].height = 21.0

    for econ_key, label, unit_txt, note_txt in econ_defs_local:
        rr = econ_rows_local[econ_key]
        build_ws.cell(row=rr, column=2, value=label)
        for start_col, end_col in ((3, 4), (5, 6), (7, 8), (9, 10), (12, 15)):
            build_ws.merge_cells(start_row=rr, start_column=start_col, end_row=rr, end_column=end_col)
        for cc in range(section_start_col, section_end_col + 1):
            build_ws.cell(row=rr, column=cc).fill = zebra_fill_dark if ((rr - subheader_row) % 2 == 0) else zebra_fill_light
            build_ws.cell(row=rr, column=cc).font = body_font
            build_ws.cell(row=rr, column=cc).border = thin_border
            build_ws.cell(row=rr, column=cc).alignment = align_left_center_wrap if cc in {2, 12} else align_center
        build_ws.cell(row=rr, column=11, value=unit_txt)
        build_ws.cell(row=rr, column=12, value=note_txt)
        for cc in (3, 5, 7, 9):
            build_ws.cell(row=rr, column=cc).number_format = "#,##0.000"
        build_ws.row_dimensions[rr].height = 24.0

    ethanol_yield_ref = source_coeff_ref.get("ethanol_yield", "$B$0")
    frame_sources = {
        "prior_quarter": source_prior_ref,
        "quarter_open": source_quarter_open_ref,
        "current_qtd": source_current_ref,
        "next_quarter_thesis": source_thesis_ref,
    }
    for frame_key, value_col in frame_cols.items():
        value_letter = get_column_letter(value_col)
        source_refs = frame_sources.get(frame_key) or {}
        build_ws.cell(
            row=econ_rows_local["ethanol_revenue"],
            column=value_col,
            value=_formula_if_ready(
                [ethanol_yield_ref, source_refs.get("ethanol_price", "")],
                f"{ethanol_yield_ref}*{source_refs.get('ethanol_price', '$B$0')}",
            ),
        )
        build_ws.cell(
            row=econ_rows_local["distillers_contribution"],
            column=value_col,
            value=_formula_if_ready(
                [source_coeff_ref.get("distillers_yield", ""), source_refs.get("distillers_grains_price", "")],
                f"{source_coeff_ref.get('distillers_yield', '$B$0')}*{source_refs.get('distillers_grains_price', '$B$0')}",
            ),
        )
        build_ws.cell(
            row=econ_rows_local["uhp_contribution"],
            column=value_col,
            value=_formula_if_ready(
                [source_coeff_ref.get("uhp_yield", ""), source_refs.get("uhp_price", "")],
                f"{source_coeff_ref.get('uhp_yield', '$B$0')}*{source_refs.get('uhp_price', '$B$0')}",
            ),
        )
        build_ws.cell(
            row=econ_rows_local["corn_oil_contribution"],
            column=value_col,
            value=_formula_if_ready(
                [source_coeff_ref.get("renewable_corn_oil_yield", ""), source_refs.get("renewable_corn_oil_price", "")],
                f"{source_coeff_ref.get('renewable_corn_oil_yield', '$B$0')}*{source_refs.get('renewable_corn_oil_price', '$B$0')}",
            ),
        )
        build_ws.cell(
            row=econ_rows_local["feedstock_cost"],
            column=value_col,
            value=_formula_if_ready(
                [source_refs.get("corn_price", "")],
                f"-{source_refs.get('corn_price', '$B$0')}",
            ),
        )
        build_ws.cell(
            row=econ_rows_local["natural_gas_burden"],
            column=value_col,
            value=_formula_if_ready(
                [source_coeff_ref.get("natural_gas_usage", ""), ethanol_yield_ref, source_refs.get("natural_gas_price", "")],
                f"-({source_coeff_ref.get('natural_gas_usage', '$B$0')}/1000000)*{ethanol_yield_ref}*{source_refs.get('natural_gas_price', '$B$0')}",
            ),
        )
        build_ws.cell(
            row=econ_rows_local["coproduct_credit"],
            column=value_col,
            value=f'=IFERROR(IF(COUNTA({value_letter}{econ_rows_local["distillers_contribution"]}:{value_letter}{econ_rows_local["corn_oil_contribution"]})=0,"",SUM({value_letter}{econ_rows_local["distillers_contribution"]}:{value_letter}{econ_rows_local["corn_oil_contribution"]})),"")',
        )
        build_ws.cell(
            row=econ_rows_local["process_margin"],
            column=value_col,
            value=_formula_if_ready(
                [ethanol_yield_ref, f"{value_letter}{econ_rows_local['ethanol_revenue']}"],
                f"({value_letter}{econ_rows_local['ethanol_revenue']}+{value_letter}{econ_rows_local['feedstock_cost']}+{value_letter}{econ_rows_local['natural_gas_burden']})/{ethanol_yield_ref}",
            ),
        )

    frame_snapshot_display_values = {
        "prior_quarter": _official_corn_basis_snapshot_display(prior_q_market_snapshot),
        "quarter_open": _official_corn_basis_snapshot_display(quarter_open_market_snapshot),
        "current_qtd": _official_corn_basis_snapshot_display(current_qtd_market_snapshot),
        "next_quarter_thesis": _official_corn_basis_snapshot_display(
            next_quarter_thesis_snapshot.get("corn") if isinstance(next_quarter_thesis_snapshot, dict) else None
        ),
    }
    frame_selection_rule_display_values = {
        "prior_quarter": _official_corn_basis_selection_rule_display(
            prior_q_market_snapshot,
            fallback_rule="latest_snapshot_on_or_before_quarter_end",
        ),
        "quarter_open": _official_corn_basis_selection_rule_display(
            quarter_open_market_snapshot,
            fallback_rule="latest_snapshot_on_or_before_quarter_start",
        ),
        "current_qtd": _official_corn_basis_selection_rule_display(
            current_qtd_market_snapshot,
            fallback_rule="latest_snapshot_on_or_before_as_of",
        ),
        "next_quarter_thesis": _official_corn_basis_selection_rule_display(
            next_quarter_thesis_snapshot.get("corn") if isinstance(next_quarter_thesis_snapshot, dict) else None,
            fallback_rule="latest_snapshot_on_or_before_as_of_with_target_quarter_rows",
        ),
    }
    for frame_key, value_col in frame_cols.items():
        build_ws.cell(
            row=econ_rows_local["corn_basis_snapshot_date"],
            column=value_col,
            value=frame_snapshot_display_values.get(frame_key) or "",
        )
        build_ws.cell(
            row=econ_rows_local["corn_basis_selection_rule"],
            column=value_col,
            value=frame_selection_rule_display_values.get(frame_key) or "",
        )

    process_margin_refs = {
        frame_key: f"{build_ws.title}!${get_column_letter(frame_cols[frame_key])}${econ_rows_local['process_margin']}"
        for frame_key in frame_cols
    }
    return {
        "title_row": title_row,
        "note_row": note_row,
        "header_row": header_row,
        "subheader_row": subheader_row,
        "econ_rows": econ_rows_local,
        "process_margin_refs": process_margin_refs,
        "next_row": econ_rows_local["corn_basis_selection_rule"],
    }




def write_gpre_economics_overlay_input_rows(
    deps: GpreEconomicsOverlayInputRowsDeps,
) -> GpreEconomicsOverlayInputRowsResult:
    ws = deps.ws
    row_idx = int(deps.row_idx)
    is_gpre_profile = deps.is_gpre_profile
    gpre_commercial_setup_rows = deps.gpre_commercial_setup_rows
    overlay_gpre_end_col = deps.overlay_gpre_end_col
    overlay_display_quarters = deps.overlay_display_quarters
    coefficient_templates = deps.coefficient_templates
    gpre_official_market_rows = deps.gpre_official_market_rows
    gpre_basis_quarter_map = deps.gpre_basis_quarter_map
    gpre_official_market_summary = deps.gpre_official_market_summary
    gpre_official_weighting_method = deps.gpre_official_weighting_method
    gpre_official_ethanol_method = deps.gpre_official_ethanol_method
    gpre_official_basis_method = deps.gpre_official_basis_method
    gpre_official_gas_method = deps.gpre_official_gas_method
    gpre_official_fallback_policy = deps.gpre_official_fallback_policy
    hidden_overlay_coefficient_keys = deps.hidden_overlay_coefficient_keys
    hidden_overlay_market_input_keys = deps.hidden_overlay_market_input_keys
    prior_market_display_quarter_txt = deps.prior_market_display_quarter_txt
    quarter_open_overlay_header_txt = deps.quarter_open_overlay_header_txt
    current_qtd_market_snapshot = deps.current_qtd_market_snapshot
    next_thesis_quarter_txt = deps.next_thesis_quarter_txt
    prior_q_market_snapshot = deps.prior_q_market_snapshot
    quarter_open_market_snapshot = deps.quarter_open_market_snapshot
    next_quarter_thesis_snapshot = deps.next_quarter_thesis_snapshot
    current_overlay_model_key = deps.current_overlay_model_key
    overlay_section_row_height = deps.overlay_section_row_height
    overlay_intro_row_height = deps.overlay_intro_row_height
    overlay_header_row_height = deps.overlay_header_row_height
    overlay_support_row_height = deps.overlay_support_row_height
    thin_border = deps.thin_border
    body_font = deps.body_font
    input_fill = deps.input_fill
    input_font = deps.input_font
    section_fill = deps.section_fill
    align_left_center_wrap = deps.align_left_center_wrap
    align_center = deps.align_center
    align_left_center = deps.align_left_center
    align_center_wrap = deps.align_center_wrap
    align_left_top_wrap = deps.align_left_top_wrap
    intro_fill = deps.intro_fill
    zebra_fill_light = deps.zebra_fill_light
    zebra_fill_dark = deps.zebra_fill_dark
    bold_font = deps.bold_font
    header_fill = deps.header_fill
    font_size = deps.font_size
    dark_text_color = deps.dark_text_color
    _write_section_bar = deps.write_section_bar
    _write_overlay_intro = deps.write_overlay_intro
    _write_header_row = deps.write_header_row
    _center_header_span = deps.center_header_span
    _write_overlay_subheader_row = deps.write_overlay_subheader_row
    _overlay_coefficient_detail = deps.overlay_coefficient_detail
    _overlay_coefficient_basis_display = deps.overlay_coefficient_basis_display
    _overlay_coefficient_source_display = deps.overlay_coefficient_source_display
    _add_comment = deps.add_comment
    _record_writer_substage = deps.record_writer_substage
    _market_input_intro_text = deps.market_input_intro_text
    _ordered_market_input_templates = deps.ordered_market_input_templates
    _pick_market_reference = deps.pick_market_reference
    _market_source_note = deps.market_source_note
    _driver_source_note = deps.driver_source_note
    _prior_market_override = deps.prior_market_override
    _quarter_open_market_override = deps.quarter_open_market_override
    _current_market_override = deps.current_market_override
    _thesis_market_override = deps.thesis_market_override
    _market_input_source_text = deps.market_input_source_text
    _overlay_as_of_header_text = deps.overlay_as_of_header_text
    _overlay_preview_bundle_for_model = deps.overlay_preview_bundle_for_model
    _snapshot_market_meta = deps.snapshot_market_meta
    market_gpre_phase_preview_story = deps.market_gpre_phase_preview_story
    coeff_rows: Dict[str, int] = {}
    market_rows: Dict[str, int] = {}
    overlay_base_coeff_started = time.perf_counter()
    row_idx = _write_section_bar(
        row_idx,
        "Base operating coefficients",
        end_col=overlay_gpre_end_col if (is_gpre_profile and gpre_commercial_setup_rows) else 5,
        primary=bool(is_gpre_profile and gpre_commercial_setup_rows),
        row_height=overlay_section_row_height if (is_gpre_profile and gpre_commercial_setup_rows) else None,
    )
    if is_gpre_profile and gpre_commercial_setup_rows:
        row_idx = _write_overlay_intro(
            row_idx,
            "Use platform/process coefficients as editable base assumptions. Reported values override inferred and user-entered assumptions when explicitly disclosed.",
            end_col=overlay_gpre_end_col,
            spacer_after=1,
            row_height=overlay_intro_row_height,
        )
        row_idx = _write_header_row(
            row_idx,
            [],
            spans=[
                (1, 1, "Coefficient"),
                (2, 2, "Base value"),
                (3, 3, "Unit"),
                (4, 5, "Status"),
                (6, 8, "Source"),
                (9, 17, "Coverage / note"),
            ],
            row_height=overlay_header_row_height,
        )
    else:
        row_idx = _write_header_row(row_idx, ["Coefficient", "Base value", "Unit", "Status", "Source"])
    for tpl in coefficient_templates:
        key = str(getattr(tpl, "key", "") or "").strip()
        coeff_detail = _overlay_coefficient_detail(key)
        value = coeff_detail.get("value")
        basis = _overlay_coefficient_basis_display(coeff_detail.get("basis"))
        source_txt = _overlay_coefficient_source_display(coeff_detail.get("source_txt"))
        source_comment = str(coeff_detail.get("source_comment") or "")
        note_txt = ""
        if key == "ethanol_yield":
            note_txt = "Report-aligned default anchored to the USDA average cited in recent GPRE filings."
        elif key == "natural_gas_usage":
            note_txt = "About 0.028 MMBtu/gal. Report-aligned process coefficient used in the GPRE overlay."
        if source_comment:
            note_txt = " ".join(part for part in (note_txt.strip(), source_comment.strip()) if part)
        ws.cell(row=row_idx, column=1, value=str(getattr(tpl, "label", "") or key))
        if is_gpre_profile and gpre_commercial_setup_rows:
            ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=5)
            ws.merge_cells(start_row=row_idx, start_column=6, end_row=row_idx, end_column=8)
            ws.merge_cells(start_row=row_idx, start_column=9, end_row=row_idx, end_column=17)
        for cc in range(1, (17 if (is_gpre_profile and gpre_commercial_setup_rows) else 5) + 1):
            ws.cell(row=row_idx, column=cc).border = thin_border
            ws.cell(row=row_idx, column=cc).alignment = (
                align_left_center_wrap if cc in {1, 4, 6, 9} else align_center if cc in {2, 3} else align_left_center
            )
            if is_gpre_profile and gpre_commercial_setup_rows:
                ws.cell(row=row_idx, column=cc).font = body_font
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        val_cell.fill = input_fill
        val_cell.font = input_font
        if value is not None:
            val_cell.number_format = "#,##0.000"
        ws.cell(row=row_idx, column=3, value=str(getattr(tpl, "unit", "") or ""))
        ws.cell(row=row_idx, column=4, value=basis)
        ws.cell(row=row_idx, column=6, value=source_txt)
        if is_gpre_profile and gpre_commercial_setup_rows:
            ws.cell(row=row_idx, column=9, value=note_txt)
        if source_comment:
            _add_comment(f"I{row_idx}" if (is_gpre_profile and gpre_commercial_setup_rows) else f"F{row_idx}", source_comment)
        if is_gpre_profile and gpre_commercial_setup_rows:
            ws.row_dimensions[row_idx].height = overlay_support_row_height
        if key in hidden_overlay_coefficient_keys:
            ws.row_dimensions[row_idx].hidden = True
        coeff_rows[key] = row_idx
        row_idx += 1

    if is_gpre_profile and gpre_commercial_setup_rows and (gpre_official_market_rows or gpre_basis_quarter_map):
        summary_text = str(gpre_official_market_summary or "").strip() or "Official market model."
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=17)
        for cc in range(1, 18):
            ws.cell(row=row_idx, column=cc).border = thin_border
            ws.cell(row=row_idx, column=cc).fill = section_fill
            ws.cell(row=row_idx, column=cc).font = Font(name=body_font.name, size=body_font.sz, bold=True, color="203864")
            ws.cell(row=row_idx, column=cc).alignment = align_left_center_wrap
        ws.cell(row=row_idx, column=1, value=summary_text)
        ws.row_dimensions[row_idx].height = max(overlay_support_row_height, 26.0)
        row_idx += 1

        method_rows = [
            ("Weighting choice", gpre_official_weighting_method),
            ("Ethanol weighting", gpre_official_ethanol_method),
            ("Basis weighting", gpre_official_basis_method),
            ("Gas burden method", gpre_official_gas_method),
            ("Fallback policy", gpre_official_fallback_policy),
        ]
        for label_txt, note_txt in method_rows:
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=17)
            for cc in range(1, 18):
                ws.cell(row=row_idx, column=cc).border = thin_border
                ws.cell(row=row_idx, column=cc).font = body_font
                ws.cell(row=row_idx, column=cc).alignment = align_left_center_wrap
            ws.cell(row=row_idx, column=1, value=label_txt)
            ws.cell(row=row_idx, column=2, value=note_txt)
            ws.row_dimensions[row_idx].height = overlay_support_row_height
            row_idx += 1

        row_idx = _write_header_row(
            row_idx,
            [],
            spans=[
                (1, 1, "Region / family"),
                (2, 2, "Capacity"),
                (3, 3, "Weight %"),
                (4, 5, "Mapped ethanol $/gal"),
                (6, 7, "Ethanol series"),
                (8, 8, "Basis c/bu"),
                (9, 9, "Basis $/bu"),
                (10, 11, "Basis series"),
                (12, 17, "Coverage / note"),
            ],
            row_height=overlay_header_row_height,
        )
        for start_col, end_col in ((4, 5), (6, 7), (10, 11), (12, 17)):
            _center_header_span(row_idx - 1, start_col, end_col)
        for rec in gpre_official_market_rows:
            ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=5)
            ws.merge_cells(start_row=row_idx, start_column=6, end_row=row_idx, end_column=7)
            ws.merge_cells(start_row=row_idx, start_column=10, end_row=row_idx, end_column=11)
            ws.merge_cells(start_row=row_idx, start_column=12, end_row=row_idx, end_column=17)
            for cc in range(1, 18):
                ws.cell(row=row_idx, column=cc).border = thin_border
                ws.cell(row=row_idx, column=cc).font = body_font
                ws.cell(row=row_idx, column=cc).alignment = (
                    align_left_center_wrap if cc in {1, 6, 10, 12} else align_center if cc in {2, 3, 4, 5, 8, 9} else align_left_center
                )
            ws.cell(row=row_idx, column=1, value=rec.get("region_label"))
            plant_cell = ws.cell(
                row=row_idx,
                column=2,
                value=rec.get("active_capacity_mmgy") if rec.get("active_capacity_mmgy") is not None else rec.get("capacity_mmgy"),
            )
            weight_cell = ws.cell(row=row_idx, column=3, value=rec.get("weight"))
            ethanol_cell = ws.cell(
                row=row_idx,
                column=4,
                value=rec.get("ethanol_value_usd_per_gal") if rec.get("ethanol_value_usd_per_gal") is not None else rec.get("ethanol_usd_per_gal"),
            )
            ws.cell(row=row_idx, column=6, value=rec.get("ethanol_series_label"))
            basis_cents_cell = ws.cell(
                row=row_idx,
                column=8,
                value=rec.get("basis_value_cents_per_bu") if rec.get("basis_value_cents_per_bu") is not None else rec.get("basis_cents_per_bu"),
            )
            basis_usd_cell = ws.cell(
                row=row_idx,
                column=9,
                value=rec.get("basis_value_usd_per_bu") if rec.get("basis_value_usd_per_bu") is not None else rec.get("basis_usd_per_bu"),
            )
            ws.cell(row=row_idx, column=10, value=rec.get("basis_series_label") or rec.get("proxy_method"))
            ws.cell(row=row_idx, column=12, value=rec.get("fallback_note") or "Primary mapped series used.")
            if pd.notna(pd.to_numeric(plant_cell.value, errors="coerce")):
                plant_cell.number_format = "0.0"
            if pd.notna(pd.to_numeric(weight_cell.value, errors="coerce")):
                weight_cell.number_format = "0%"
            if pd.notna(pd.to_numeric(ethanol_cell.value, errors="coerce")):
                ethanol_cell.number_format = "0.000"
            if pd.notna(pd.to_numeric(basis_cents_cell.value, errors="coerce")):
                basis_cents_cell.number_format = "0.0"
            if pd.notna(pd.to_numeric(basis_usd_cell.value, errors="coerce")):
                basis_usd_cell.number_format = "0.000"
            ws.row_dimensions[row_idx].height = overlay_support_row_height
            row_idx += 1

    if not (is_gpre_profile and gpre_commercial_setup_rows):
        note_row = row_idx
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
        ws.cell(row=note_row, column=1, value="Use platform/process coefficients as editable base assumptions. Reported values override inferred and user-entered assumptions when explicitly disclosed.")
        ws.cell(row=note_row, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(row=note_row, column=1).border = thin_border
        ws.row_dimensions[note_row].height = 30
        row_idx += 2
    else:
        row_idx += 1
    _record_writer_substage("write_excel.drivers.render.economics_overlay.base_coefficients", overlay_base_coeff_started)

    overlay_market_inputs_started = time.perf_counter()
    market_section_bar_row = row_idx
    row_idx = _write_section_bar(
        row_idx,
        "Market inputs",
        end_col=overlay_gpre_end_col if (is_gpre_profile and gpre_commercial_setup_rows) else 5,
        primary=bool(is_gpre_profile and gpre_commercial_setup_rows),
        row_height=overlay_section_row_height if (is_gpre_profile and gpre_commercial_setup_rows) else None,
    )
    row_idx = _write_overlay_intro(
        row_idx,
        _market_input_intro_text(),
        end_col=overlay_gpre_end_col if (is_gpre_profile and gpre_commercial_setup_rows) else 5,
        spacer_after=1,
        row_height=18.0 if (is_gpre_profile and gpre_commercial_setup_rows) else 30.0,
    )
    if is_gpre_profile and gpre_commercial_setup_rows:
        row_idx = _write_header_row(
            row_idx,
            [],
            spans=[
                (1, 1, "Input"),
                (2, 3, "Prior quarter"),
                (4, 5, "Quarter-open outlook"),
                (6, 7, "Current QTD"),
                (8, 9, "Next quarter outlook"),
                (10, 10, "Unit"),
                (11, 21, "Source"),
            ],
            row_height=overlay_header_row_height,
        )
        for start_col, end_col in ((2, 3), (4, 5), (6, 7), (8, 9), (10, 10), (11, 21)):
            _center_header_span(row_idx - 1, start_col, end_col)
        row_idx = _write_overlay_subheader_row(
            row_idx,
            prior_txt=prior_market_display_quarter_txt,
            quarter_open_txt=quarter_open_overlay_header_txt,
            current_txt=_overlay_as_of_header_text(current_qtd_market_snapshot.get("as_of") if isinstance(current_qtd_market_snapshot, dict) else None),
            thesis_txt=next_thesis_quarter_txt,
            note_start_col=11,
            note_end_col=21,
            row_height=21.0,
        )
    else:
        row_idx = _write_header_row(row_idx, ["Input", "Current QTD", "Next quarter outlook", "Unit", "Source note"])
    for tpl in _ordered_market_input_templates():
        key = str(getattr(tpl, "key", "") or "").strip()
        market_ref = _pick_market_reference(tpl)
        current_val = market_ref.get("_converted_value") if market_ref else None
        source_txt = _market_source_note(market_ref)
        source_comment = _driver_source_note(
            market_ref.get("source_file"),

            market_ref.get("parsed_text"),

            f"series={market_ref.get('series_key')} | region={market_ref.get('region')} | unit={market_ref.get('unit')}",

        ) if market_ref else ""
        if key == "uhp_price" and not source_txt:
            source_txt = "User assumption"
            if not source_comment:
                source_comment = "No explicit quarterly market quote selected; manual ultra-high protein price input remains assumption-driven."
        prior_override = _prior_market_override(key)
        quarter_open_override = _quarter_open_market_override(key)
        current_override = _current_market_override(key)
        thesis_override = _thesis_market_override(key)
        if isinstance(current_override, dict):
            current_val = current_override.get("value")
        prior_val = (prior_override or {}).get("value") if isinstance(prior_override, dict) else None
        quarter_open_val = (quarter_open_override or {}).get("value") if isinstance(quarter_open_override, dict) else None
        source_comment_parts = [
            part
            for part in (
                str((prior_override or {}).get("comment") or "").strip(),
                str((quarter_open_override or {}).get("comment") or "").strip(),
                str((current_override or {}).get("comment") or "").strip(),
            )
            if part
        ]
        if source_comment_parts:
            source_comment = " ".join(source_comment_parts)
        if isinstance(thesis_override, dict) and str(thesis_override.get("comment") or "").strip():
            source_comment = " ".join(
                part
                for part in (
                    str(source_comment or "").strip(),
                    str(thesis_override.get("comment") or "").strip(),
                )
                if part
            )
        source_display_txt = _market_input_source_text(
            key,
            prior_override,
            quarter_open_override,
            current_override,
            thesis_override,
            source_txt,
        )
        ws.cell(row=row_idx, column=1, value=str(getattr(tpl, "label", "") or key))
        if is_gpre_profile and gpre_commercial_setup_rows:
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
            ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=5)
            ws.merge_cells(start_row=row_idx, start_column=6, end_row=row_idx, end_column=7)
            ws.merge_cells(start_row=row_idx, start_column=8, end_row=row_idx, end_column=9)
            ws.merge_cells(start_row=row_idx, start_column=11, end_row=row_idx, end_column=21)
        market_wrap_cols = {1, 11} if (is_gpre_profile and gpre_commercial_setup_rows) else {1, 5}
        for cc in range(1, ((overlay_gpre_end_col if (is_gpre_profile and gpre_commercial_setup_rows) else 5)) + 1):
            ws.cell(row=row_idx, column=cc).border = thin_border
            ws.cell(row=row_idx, column=cc).alignment = Alignment(horizontal="left" if cc not in {2, 3, 4, 5, 6, 7, 8, 9, 10} else "center", vertical="center", wrap_text=cc in market_wrap_cols)
            if is_gpre_profile and gpre_commercial_setup_rows:
                ws.cell(row=row_idx, column=cc).font = body_font
        prior_cell = ws.cell(row=row_idx, column=2, value=prior_val)
        quarter_open_cell = ws.cell(row=row_idx, column=4, value=quarter_open_val)
        current_cell = ws.cell(row=row_idx, column=6, value=current_val)
        if prior_val is not None:
            prior_cell.number_format = "#,##0.000"
        if quarter_open_val is not None:
            quarter_open_cell.number_format = "#,##0.000"
        if current_val is not None:
            current_cell.number_format = "#,##0.000"
        thesis_col = 8 if (is_gpre_profile and gpre_commercial_setup_rows) else 3
        note_col = 11 if (is_gpre_profile and gpre_commercial_setup_rows) else 5
        unit_col = 10 if (is_gpre_profile and gpre_commercial_setup_rows) else 4
        thesis_cell = ws.cell(row=row_idx, column=thesis_col)
        proxy_base_key = str(getattr(tpl, "proxy_base_key", "") or "").strip()
        proxy_premium_key = str(getattr(tpl, "proxy_premium_key", "") or "").strip()
        if proxy_base_key and proxy_premium_key:
            base_row = market_rows.get(proxy_base_key)
            prem_row = market_rows.get(proxy_premium_key)
            if base_row and prem_row:
                thesis_ref_letter = get_column_letter(thesis_col)
                thesis_cell.value = f'=IF(AND(ISNUMBER({thesis_ref_letter}{base_row}),ISNUMBER({thesis_ref_letter}{prem_row})),{thesis_ref_letter}{base_row}+{thesis_ref_letter}{prem_row},IF(AND(ISNUMBER(B{base_row}),ISNUMBER({thesis_ref_letter}{prem_row})),B{base_row}+{thesis_ref_letter}{prem_row},""))'
                prior_cell.value = f'=IF(AND(ISNUMBER(B{base_row}),ISNUMBER(B{prem_row})),B{base_row}+B{prem_row},"")'
                quarter_open_cell.value = f'=IF(AND(ISNUMBER(D{base_row}),ISNUMBER(D{prem_row})),D{base_row}+D{prem_row},"")'
                current_cell.value = f'=IF(AND(ISNUMBER(F{base_row}),ISNUMBER(F{prem_row})),F{base_row}+F{prem_row},"")'
                source_txt = "Proxy from soybean oil + premium"
                source_comment = "Proxy row: direct external soybean oil reference plus manual corn-oil premium assumption."
        elif key == "corn_oil_premium_assumption":
            thesis_cell.fill = input_fill
            thesis_cell.font = input_font
            thesis_cell.number_format = "#,##0.000"
        elif isinstance(thesis_override, dict) and thesis_override.get("manual") is False:
            thesis_cell.value = thesis_override.get("value")
            thesis_cell.number_format = "#,##0.000"
        else:
            thesis_cell.fill = input_fill
            thesis_cell.font = input_font
        if is_gpre_profile and gpre_commercial_setup_rows:
            for cc in range(thesis_col, thesis_col + 2):
                ws.cell(row=row_idx, column=cc).fill = copy(thesis_cell.fill)
                ws.cell(row=row_idx, column=cc).font = copy(thesis_cell.font)
        ws.cell(row=row_idx, column=unit_col, value=str(getattr(tpl, "unit", "") or ""))
        meta_col = 11 if (is_gpre_profile and gpre_commercial_setup_rows) else 5
        ws.cell(row=row_idx, column=meta_col, value=source_display_txt if is_gpre_profile and gpre_commercial_setup_rows else source_txt)
        if source_comment:
            _add_comment(f"{get_column_letter(meta_col)}{row_idx}", source_comment)
        if is_gpre_profile and gpre_commercial_setup_rows:
            ws.row_dimensions[row_idx].height = overlay_support_row_height
        if key in hidden_overlay_market_input_keys:
            ws.row_dimensions[row_idx].hidden = True
        market_rows[key] = row_idx
        row_idx += 1

    corn_oil_row = market_rows.get("renewable_corn_oil_price")
    implied_proxy_row = market_rows.get("implied_renewable_corn_oil_proxy_price")
    if corn_oil_row and implied_proxy_row:
        prior_cell = ws.cell(row=corn_oil_row, column=2)
        quarter_open_cell = ws.cell(row=corn_oil_row, column=(4 if (is_gpre_profile and gpre_commercial_setup_rows) else 2))
        current_cell = ws.cell(row=corn_oil_row, column=(6 if (is_gpre_profile and gpre_commercial_setup_rows) else 2))
        if prior_cell.value in (None, ""):
            prior_cell.value = f'=IF(ISNUMBER(B{implied_proxy_row}),B{implied_proxy_row},"")'
        if is_gpre_profile and gpre_commercial_setup_rows and quarter_open_cell.value in (None, ""):
            quarter_open_cell.value = f'=IF(ISNUMBER(D{implied_proxy_row}),D{implied_proxy_row},"")'
        if is_gpre_profile and gpre_commercial_setup_rows and current_cell.value in (None, ""):
            current_cell.value = f'=IF(ISNUMBER(F{implied_proxy_row}),F{implied_proxy_row},"")'
            note_col = 11 if (is_gpre_profile and gpre_commercial_setup_rows) else 5
            if not str(ws.cell(row=corn_oil_row, column=note_col).value or "").strip():
                ws.cell(row=corn_oil_row, column=note_col, value="Proxy from soybean oil + premium")
                _add_comment(
                    f"{get_column_letter(note_col)}{corn_oil_row}",
                    "Direct renewable corn oil market reference unavailable; external reference is proxied from soybean oil plus the manual corn-oil premium assumption.",
                )

    row_idx += 1
    coeff_ref = {k: f"$B${r}" for k, r in coeff_rows.items()}
    prior_ref = {k: f"$B${r}" for k, r in market_rows.items()}
    quarter_open_ref = {k: f"${'D' if (is_gpre_profile and gpre_commercial_setup_rows) else 'B'}${r}" for k, r in market_rows.items()}
    current_ref = {k: f"${'F' if (is_gpre_profile and gpre_commercial_setup_rows) else 'B'}${r}" for k, r in market_rows.items()}
    thesis_ref = {k: f"${'H' if (is_gpre_profile and gpre_commercial_setup_rows) else 'C'}${r}" for k, r in market_rows.items()}


    def _formula_if_ready(refs: List[str], expr: str) -> str:
        use_refs = [ref for ref in refs if ref]
        if not use_refs:
            return '=""'
        checks = ",".join([f"ISNUMBER({ref})" for ref in use_refs])
        return f'=IFERROR(IF(AND({checks}),{expr},""),"")'

    def _gpre_model_formula_helper(model_key: str, phase_key: str) -> Dict[str, Any]:
        helper = (((_overlay_preview_bundle_for_model(model_key) or {}).get("gpre_proxy_formula_helpers") or {}).get(phase_key) or {})
        return dict(helper) if isinstance(helper, dict) else {}

    def _gpre_formula_helper(phase_key: str) -> Dict[str, Any]:
        return _gpre_model_formula_helper(current_overlay_model_key, phase_key)

    def _gpre_model_formula_note(model_key: str, phase_key: str) -> str:
        helper = _gpre_model_formula_helper(model_key, phase_key)
        helper_note = str(helper.get("live_preview_note") or "").strip()
        if helper_note:
            return helper_note
        preview_story = market_gpre_phase_preview_story(
            str(model_key or current_overlay_model_key),
            phase=(
                "current"
                if str(phase_key or "") == "current_qtd"
                else "next"
                if str(phase_key or "") == "next_quarter_thesis"
                else str(phase_key or "")
            ),
        )
        note_txt = str((preview_story or {}).get("live_preview_note") or "").strip()
        if note_txt:
            return note_txt
        if str(phase_key or "") == "quarter_open":
            return "Quarter-open fitted value for the chosen model."
        return ""

    def _gpre_formula_note(phase_key: str) -> str:
        return _gpre_model_formula_note(current_overlay_model_key, phase_key)

    def _gpre_model_live_formula(model_key: str, phase_key: str, ethanol_ref: str) -> Optional[str]:
        helper = _gpre_model_formula_helper(model_key, phase_key)
        slope_num = pd.to_numeric(helper.get("slope"), errors="coerce")
        intercept_num = pd.to_numeric(helper.get("intercept"), errors="coerce")
        if pd.isna(slope_num) or pd.isna(intercept_num):
            return None
        if abs(float(slope_num)) <= 1e-12:
            return f"={float(intercept_num):.12f}"
        if not ethanol_ref:
            return None
        expr = f"({float(slope_num):.12f}*{ethanol_ref})+({float(intercept_num):.12f})"
        return _formula_if_ready([ethanol_ref], expr)

    def _gpre_fitted_live_formula(phase_key: str, ethanol_ref: str) -> Optional[str]:
        return _gpre_model_live_formula(current_overlay_model_key, phase_key, ethanol_ref)

    def _sheet_ref(ref_in: str, source_sheet_name: str = "") -> str:
        ref_txt = str(ref_in or "").strip()
        if not ref_txt:
            return ""
        return f"{source_sheet_name}!{ref_txt}" if source_sheet_name else ref_txt

    def _official_corn_basis_snapshot_display(snapshot_in: Optional[Dict[str, Any]]) -> str:
        meta = _snapshot_market_meta(snapshot_in, "corn_price")
        snapshot_date_val = (
            meta.get("official_corn_basis_snapshot_date")
            if isinstance(meta, dict)
            else None
        )
        if not isinstance(snapshot_date_val, date) and isinstance(snapshot_in, dict):
            snapshot_date_val = (
                snapshot_in.get("official_corn_basis_snapshot_date")
                or snapshot_in.get("snapshot_date")
            )
        if not isinstance(snapshot_date_val, date):
            parsed_snapshot_date = pd.to_datetime(snapshot_date_val, errors="coerce")
            if pd.notna(parsed_snapshot_date):
                snapshot_date_val = pd.Timestamp(parsed_snapshot_date).date()
        if isinstance(snapshot_date_val, date):
            return snapshot_date_val.isoformat()
        source_kind_txt = (
            str(meta.get("official_corn_basis_source_kind") or "").strip()
            if isinstance(meta, dict)
            else ""
        )
        if not source_kind_txt and isinstance(snapshot_in, dict):
            source_kind_txt = str(snapshot_in.get("official_corn_basis_source_kind") or "").strip()
        if source_kind_txt == "weighted_ams_proxy":
            return "AMS fallback"
        return ""

    def _official_corn_basis_selection_rule_display(
        snapshot_in: Optional[Dict[str, Any]],
        *,
        fallback_rule: str,
    ) -> str:
        meta = _snapshot_market_meta(snapshot_in, "corn_price")
        selection_rule_txt = (
            str(meta.get("official_corn_basis_selection_rule") or "").strip()
            if isinstance(meta, dict)
            else ""
        )
        if not selection_rule_txt and isinstance(snapshot_in, dict):
            selection_rule_txt = str(
                snapshot_in.get("official_corn_basis_selection_rule")
                or snapshot_in.get("selection_rule")
                or ""
            ).strip()
        if not selection_rule_txt:
            selection_rule_txt = str(fallback_rule or "").strip()
        source_kind_txt = (
            str(meta.get("official_corn_basis_source_kind") or "").strip()
            if isinstance(meta, dict)
            else ""
        )
        if not source_kind_txt and isinstance(snapshot_in, dict):
            source_kind_txt = str(snapshot_in.get("official_corn_basis_source_kind") or "").strip()
        if source_kind_txt == "weighted_ams_proxy":
            return f"{selection_rule_txt} / AMS fallback" if selection_rule_txt else "AMS fallback"
        return selection_rule_txt


    build_up_deps = GpreEconomicsOverlayBuildUpDeps(
        coeff_rows=dict(coeff_rows),
        market_rows=dict(market_rows),
        section_fill=section_fill,
        bold_font=bold_font,
        align_center=align_center,
        thin_border=thin_border,
        intro_fill=intro_fill,
        body_font=body_font,
        align_left_top_wrap=align_left_top_wrap,
        header_fill=header_fill,
        align_center_wrap=align_center_wrap,
        zebra_fill_light=zebra_fill_light,
        zebra_fill_dark=zebra_fill_dark,
        align_left_center_wrap=align_left_center_wrap,
        prior_market_display_quarter_txt=prior_market_display_quarter_txt,
        quarter_open_overlay_header_txt=quarter_open_overlay_header_txt,
        current_qtd_market_snapshot=current_qtd_market_snapshot,
        next_thesis_quarter_txt=next_thesis_quarter_txt,
        prior_q_market_snapshot=prior_q_market_snapshot,
        quarter_open_market_snapshot=quarter_open_market_snapshot,
        next_quarter_thesis_snapshot=next_quarter_thesis_snapshot,
        overlay_as_of_header_text=_overlay_as_of_header_text,
        official_corn_basis_snapshot_display=_official_corn_basis_snapshot_display,
        official_corn_basis_selection_rule_display=_official_corn_basis_selection_rule_display,
    )

    def _write_gpre_approx_market_crush_build_up_section(
        target_ws: Any,
        start_row_in: int,
        *,
        source_sheet_name: str,
    ) -> Dict[str, Any]:
        return write_gpre_approx_market_crush_build_up_section(
            build_up_deps,
            target_ws,
            start_row_in,
            source_sheet_name=source_sheet_name,
        )

    return GpreEconomicsOverlayInputRowsResult(
        row_idx=row_idx,
        coeff_rows=dict(coeff_rows),
        market_rows=dict(market_rows),
        coeff_ref=dict(coeff_ref),
        prior_ref=dict(prior_ref),
        quarter_open_ref=dict(quarter_open_ref),
        current_ref=dict(current_ref),
        thesis_ref=dict(thesis_ref),
        market_section_bar_row=market_section_bar_row,
        market_inputs_started=overlay_market_inputs_started,
        write_gpre_approx_market_crush_build_up_section=_write_gpre_approx_market_crush_build_up_section,
        gpre_fitted_live_formula=_gpre_fitted_live_formula,
        gpre_formula_note=_gpre_formula_note,
        gpre_model_live_formula=_gpre_model_live_formula,
        gpre_model_formula_note=_gpre_model_formula_note,
    )
