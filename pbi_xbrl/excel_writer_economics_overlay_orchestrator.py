"""Orchestrates Economics_Overlay sheet rendering."""
from __future__ import annotations

import re
import time
from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side


@dataclass(frozen=True)
class EconomicsOverlayOrchestratorDeps:
    BasisProxySandboxWriterDeps: Any
    EconomicsOverlayChartWriterDeps: Any
    EconomicsOverlayMarketStateDeps: Any
    EconomicsOverlaySourceSupport: Any
    EconomicsOverlaySourceSupportDeps: Any
    GpreEconomicsOverlayBridgeDeps: Any
    GpreEconomicsOverlayCommercialDeps: Any
    GpreEconomicsOverlayCoproductDeps: Any
    GpreEconomicsOverlayCurrentQtdDeps: Any
    GpreEconomicsOverlayDerivativeSideEffectDeps: Any
    GpreEconomicsOverlayInputRowsDeps: Any
    GpreOverlayQuarterComparisonDeps: Any
    GpreOverlaySupportInputs: Any
    _apply_chart_text_categories: Any
    _convert_market_price_value: Any
    _driver_source_display: Any
    _driver_source_note: Any
    _economics_market_region_tags: Any
    _ensure_terminal_period: Any
    _estimate_wrapped_row_height: Any
    _extract_operating_driver_rows_for_template: Any
    _get_analysis_sheet_style_bundle: Any
    _gpre_commercial_setup_records_shared: Any
    _gpre_parse_snapshot_date_like: Any
    _load_operating_driver_bridge_bundle_map: Any
    _load_operating_driver_flat_line_index: Any
    _load_operating_driver_source_records_by_quarter: Any
    _load_operating_driver_template_index: Any
    _operating_driver_quarters: Any
    _overlay_model_label: Any
    _parse_driver_number: Any
    _quarter_label_short: Any
    _record_writer_substage: Any
    _set_cell_comment_local: Any
    _text_fragment_penalty: Any
    _truncate_driver_text: Any
    _write_derivative_crush_tests_sheet: Any
    build_current_qtd_simple_crush_snapshot: Any
    build_derivative_crush_tests: Any
    build_economics_overlay_market_state: Any
    build_gpre_basis_proxy_model: Any
    build_gpre_official_proxy_history_series: Any
    build_gpre_official_proxy_snapshot: Any
    build_gpre_overlay_proxy_preview_bundle: Any
    build_gpre_plant_capacity_history: Any
    build_next_quarter_thesis_snapshot: Any
    build_prior_quarter_simple_crush_snapshot: Any
    build_simple_crush_history_series: Any
    cache_dir: Any
    company_profile: Any
    data_root_from_sec_cache_path: Any
    derivative_oci_bridge_df: Any
    derivative_oci_exposure_df: Any
    economics_market_rows: Any
    fetch_gpre_corn_bids_snapshot: Any
    font_size: Any
    glx_normalize_text: Any
    header_size: Any
    info_log: Any
    is_gpre_profile: Any
    derivative_crush_tests_owned: Any
    is_pbi_profile: Any
    load_or_download_gpre_corn_bids_snapshot: Any
    market_build_gpre_proxy_implied_results_bundle: Any
    market_gpre_phase_preview_story: Any
    market_input_fingerprint: Any
    operating_driver_history_rows: Any
    persist_gpre_frozen_thesis_snapshot: Any
    qn_is_complete_signal_text: Any
    resolve_gpre_quarter_open_snapshot: Any
    state: Any
    ticker: Any
    ticker_roots: Any
    wb: Any
    write_basis_proxy_sandbox_sheet: Any
    write_economics_overlay_charts: Any
    write_gpre_basis_proxy_overlay_support: Any
    write_gpre_derivative_crush_tests_side_effect: Any
    write_gpre_economics_overlay_bridge_to_reported_section: Any
    write_gpre_economics_overlay_commercial_sections: Any
    write_gpre_economics_overlay_coproduct_section: Any
    write_gpre_economics_overlay_current_qtd_section: Any
    write_gpre_economics_overlay_input_rows: Any
    write_gpre_overlay_quarter_comparisons: Any



def write_economics_overlay_sheet(
    deps: EconomicsOverlayOrchestratorDeps,
    rows: Sequence[Mapping[str, Any]],
) -> None:
    BasisProxySandboxWriterDeps = deps.BasisProxySandboxWriterDeps
    EconomicsOverlayChartWriterDeps = deps.EconomicsOverlayChartWriterDeps
    EconomicsOverlayMarketStateDeps = deps.EconomicsOverlayMarketStateDeps
    EconomicsOverlaySourceSupport = deps.EconomicsOverlaySourceSupport
    EconomicsOverlaySourceSupportDeps = deps.EconomicsOverlaySourceSupportDeps
    GpreEconomicsOverlayBridgeDeps = deps.GpreEconomicsOverlayBridgeDeps
    GpreEconomicsOverlayCommercialDeps = deps.GpreEconomicsOverlayCommercialDeps
    GpreEconomicsOverlayCoproductDeps = deps.GpreEconomicsOverlayCoproductDeps
    GpreEconomicsOverlayCurrentQtdDeps = deps.GpreEconomicsOverlayCurrentQtdDeps
    GpreEconomicsOverlayDerivativeSideEffectDeps = deps.GpreEconomicsOverlayDerivativeSideEffectDeps
    GpreEconomicsOverlayInputRowsDeps = deps.GpreEconomicsOverlayInputRowsDeps
    GpreOverlayQuarterComparisonDeps = deps.GpreOverlayQuarterComparisonDeps
    GpreOverlaySupportInputs = deps.GpreOverlaySupportInputs
    _apply_chart_text_categories = deps._apply_chart_text_categories
    _convert_market_price_value = deps._convert_market_price_value
    _driver_source_display = deps._driver_source_display
    _driver_source_note = deps._driver_source_note
    _economics_market_region_tags = deps._economics_market_region_tags
    _ensure_terminal_period = deps._ensure_terminal_period
    _estimate_wrapped_row_height = deps._estimate_wrapped_row_height
    _extract_operating_driver_rows_for_template = deps._extract_operating_driver_rows_for_template
    _get_analysis_sheet_style_bundle = deps._get_analysis_sheet_style_bundle
    _gpre_commercial_setup_records_shared = deps._gpre_commercial_setup_records_shared
    _gpre_parse_snapshot_date_like = deps._gpre_parse_snapshot_date_like
    _load_operating_driver_bridge_bundle_map = deps._load_operating_driver_bridge_bundle_map
    _load_operating_driver_flat_line_index = deps._load_operating_driver_flat_line_index
    _load_operating_driver_source_records_by_quarter = deps._load_operating_driver_source_records_by_quarter
    _load_operating_driver_template_index = deps._load_operating_driver_template_index
    _operating_driver_quarters = deps._operating_driver_quarters
    _overlay_model_label = deps._overlay_model_label
    _parse_driver_number = deps._parse_driver_number
    _quarter_label_short = deps._quarter_label_short
    _record_writer_substage = deps._record_writer_substage
    _set_cell_comment_local = deps._set_cell_comment_local
    _text_fragment_penalty = deps._text_fragment_penalty
    _truncate_driver_text = deps._truncate_driver_text
    _write_derivative_crush_tests_sheet = deps._write_derivative_crush_tests_sheet
    build_current_qtd_simple_crush_snapshot = deps.build_current_qtd_simple_crush_snapshot
    build_derivative_crush_tests = deps.build_derivative_crush_tests
    build_economics_overlay_market_state = deps.build_economics_overlay_market_state
    build_gpre_basis_proxy_model = deps.build_gpre_basis_proxy_model
    build_gpre_official_proxy_history_series = deps.build_gpre_official_proxy_history_series
    build_gpre_official_proxy_snapshot = deps.build_gpre_official_proxy_snapshot
    build_gpre_overlay_proxy_preview_bundle = deps.build_gpre_overlay_proxy_preview_bundle
    build_gpre_plant_capacity_history = deps.build_gpre_plant_capacity_history
    build_next_quarter_thesis_snapshot = deps.build_next_quarter_thesis_snapshot
    build_prior_quarter_simple_crush_snapshot = deps.build_prior_quarter_simple_crush_snapshot
    build_simple_crush_history_series = deps.build_simple_crush_history_series
    cache_dir = deps.cache_dir
    company_profile = deps.company_profile
    data_root_from_sec_cache_path = deps.data_root_from_sec_cache_path
    derivative_oci_bridge_df = deps.derivative_oci_bridge_df
    derivative_oci_exposure_df = deps.derivative_oci_exposure_df
    economics_market_rows = deps.economics_market_rows
    fetch_gpre_corn_bids_snapshot = deps.fetch_gpre_corn_bids_snapshot
    font_size = deps.font_size
    glx_normalize_text = deps.glx_normalize_text
    header_size = deps.header_size
    info_log = deps.info_log
    is_gpre_profile = deps.is_gpre_profile
    derivative_crush_tests_owned = bool(deps.derivative_crush_tests_owned)
    is_pbi_profile = deps.is_pbi_profile
    load_or_download_gpre_corn_bids_snapshot = deps.load_or_download_gpre_corn_bids_snapshot
    market_build_gpre_proxy_implied_results_bundle = deps.market_build_gpre_proxy_implied_results_bundle
    market_gpre_phase_preview_story = deps.market_gpre_phase_preview_story
    market_input_fingerprint = deps.market_input_fingerprint
    operating_driver_history_rows = deps.operating_driver_history_rows
    persist_gpre_frozen_thesis_snapshot = deps.persist_gpre_frozen_thesis_snapshot
    qn_is_complete_signal_text = deps.qn_is_complete_signal_text
    resolve_gpre_quarter_open_snapshot = deps.resolve_gpre_quarter_open_snapshot
    state = deps.state
    ticker = deps.ticker
    ticker_roots = deps.ticker_roots
    wb = deps.wb
    write_basis_proxy_sandbox_sheet = deps.write_basis_proxy_sandbox_sheet
    write_economics_overlay_charts = deps.write_economics_overlay_charts
    write_gpre_basis_proxy_overlay_support = deps.write_gpre_basis_proxy_overlay_support
    write_gpre_derivative_crush_tests_side_effect = deps.write_gpre_derivative_crush_tests_side_effect
    write_gpre_economics_overlay_bridge_to_reported_section = deps.write_gpre_economics_overlay_bridge_to_reported_section
    write_gpre_economics_overlay_commercial_sections = deps.write_gpre_economics_overlay_commercial_sections
    write_gpre_economics_overlay_coproduct_section = deps.write_gpre_economics_overlay_coproduct_section
    write_gpre_economics_overlay_current_qtd_section = deps.write_gpre_economics_overlay_current_qtd_section
    write_gpre_economics_overlay_input_rows = deps.write_gpre_economics_overlay_input_rows
    write_gpre_overlay_quarter_comparisons = deps.write_gpre_overlay_quarter_comparisons
    overlay_setup_started = time.perf_counter()
    ws = wb.create_sheet("Economics_Overlay")
    ws.sheet_format.defaultRowHeight = 18
    ws.sheet_view.zoomScale = 110
    ws.freeze_panes = "B2"

    analysis_theme = _get_analysis_sheet_style_bundle()
    border_color = str(analysis_theme["border_color"])
    dark_text_color = str(analysis_theme["text_dark"])
    muted_text_color = str(analysis_theme["text_muted"])
    accent_text_color = str(analysis_theme["accent_text"])
    thin = copy(analysis_theme["thin_side"])
    thin_border = copy(analysis_theme["thin_border"])
    row_border = Border(bottom=thin)
    title_fill = copy(analysis_theme["title_fill"])
    section_fill = copy(analysis_theme["section_fill"])
    hero_fill = copy(analysis_theme["title_fill"])
    header_fill = copy(analysis_theme["header_fill"])
    year_band_fill = copy(analysis_theme["section_fill"])
    intro_fill = copy(analysis_theme["section_fill"])
    zebra_fill_light = copy(analysis_theme["neutral_fill_alt"])
    zebra_fill_dark = copy(analysis_theme["neutral_fill"])
    input_fill = copy(analysis_theme["input_fill"])
    title_font = Font(bold=True, size=16, color="FFFFFF")
    bold_font = Font(bold=True, size=header_size, color=dark_text_color)
    hero_font = Font(bold=True, size=header_size, color="FFFFFF")
    support_section_font = Font(bold=True, size=header_size, color=dark_text_color)
    year_band_font = Font(bold=True, size=header_size, color=accent_text_color)
    horizon_font = Font(bold=True, size=font_size, color=accent_text_color)
    stated_font = Font(size=font_size, color=muted_text_color)
    setup_font = Font(bold=True, size=font_size, color=dark_text_color)
    body_font = Font(size=font_size, color=dark_text_color)
    input_font = Font(color="0000FF", size=font_size, bold=False)
    align_center = Alignment(horizontal="center", vertical="center")
    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_center = Alignment(horizontal="left", vertical="center")
    align_left_center_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_left_top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    overlay_section_row_height = 22.5
    overlay_intro_row_height = 24.0
    overlay_header_row_height = 21.0
    overlay_year_band_row_height = 21.0
    overlay_support_row_height = 24.0
    overlay_commentary_section_row_height = 22.5
    overlay_commentary_header_row_height = 21.0
    overlay_commentary_year_band_row_height = 21.0
    overlay_commercial_section_row_height = 22.5
    overlay_commercial_header_row_height = 21.0
    overlay_commercial_year_band_row_height = 21.0
    overlay_commercial_row_max_height = 120.0
    overlay_spacer_row_height = 15.0
    quarter_separator_side = Side(style="thin", color="B8CCE4")

    def _px_to_width(px: float) -> float:
        try:
            p = float(px)
        except Exception:
            p = 100.0
        return max(1.0, round((p - 5.0) / 7.0, 2))

    def _driver_display_quarters() -> List[date]:
        quarter_pool = sorted(
            qd for qd in _operating_driver_quarters() if any(r.get("Quarter") == qd for r in rows)
        )
        return quarter_pool[-12:] if len(quarter_pool) > 12 else quarter_pool

    def _overlay_driver_source_priority(source_type_in: Any) -> int:
        source_type_txt = str(source_type_in or "").strip().lower()
        return (
            0 if source_type_txt == "earnings_release"
            else 1 if source_type_txt == "presentation"
            else 2 if source_type_txt == "press_release"
            else 3 if source_type_txt in {"10-q", "10-k"}
            else 4 if source_type_txt == "transcript"
            else 5
        )

    def _add_comment(cell_ref: str, text_in: Any) -> None:
        try:
            _set_cell_comment_local(ws[cell_ref], text_in)
        except Exception:
            pass

    def _write_section_bar(
        row_num: int,
        title: str,
        end_col: int = 8,
        *,
        primary: bool = False,
        row_height: Optional[float] = None,
    ) -> int:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=end_col)
        cell = ws.cell(row=row_num, column=1, value=title)
        cell.font = hero_font if primary else support_section_font
        cell.fill = hero_fill if primary else section_fill
        cell.alignment = align_center
        for cc in range(1, end_col + 1):
            ws.cell(row=row_num, column=cc).fill = hero_fill if primary else section_fill
            ws.cell(row=row_num, column=cc).border = thin_border
        ws.row_dimensions[row_num].height = float(row_height if row_height is not None else (20 if primary else 18))
        return row_num + 1

    def _write_header_row(
        row_num: int,
        headers: List[str],
        *,
        spans: Optional[List[Tuple[int, int, str]]] = None,
        row_height: Optional[float] = None,
    ) -> int:
        use_spans = spans or [(idx, idx, label) for idx, label in enumerate(headers, start=1)]
        for start_col, end_col, label in use_spans:
            if end_col > start_col:
                ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
            cell = ws.cell(row=row_num, column=start_col, value=label)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = align_left_center_wrap
            cell.border = thin_border
            for cc in range(start_col, end_col + 1):
                ws.cell(row=row_num, column=cc).fill = header_fill
                ws.cell(row=row_num, column=cc).border = thin_border
                ws.cell(row=row_num, column=cc).alignment = align_left_center_wrap
        ws.row_dimensions[row_num].height = float(row_height if row_height is not None else 20)
        return row_num + 1

    source_lines = _load_operating_driver_flat_line_index()

    derivative_bridge_by_quarter: Dict[date, Dict[str, Any]] = {}
    if isinstance(derivative_oci_bridge_df, pd.DataFrame) and not derivative_oci_bridge_df.empty:
        for _, der_row in derivative_oci_bridge_df.iterrows():
            der_q = pd.to_datetime(der_row.get("quarter"), errors="coerce")
            if pd.isna(der_q):
                continue
            derivative_bridge_by_quarter[der_q.date()] = dict(der_row)


    quarter_set = _driver_display_quarters()
    as_of_market_quarter = max(quarter_set) if quarter_set else None
    overlay_market_as_of = date.today()

    market_input_templates = list(getattr(company_profile, "economics_overlay_market_inputs", ()) or [])
    market_input_templates_by_key = {
        str(getattr(tpl, "key", "") or "").strip(): tpl
        for tpl in market_input_templates
        if str(getattr(tpl, "key", "") or "").strip()
    }
    coefficient_templates = list(getattr(company_profile, "economics_overlay_coefficients", ()) or [])
    hidden_overlay_coefficient_keys = {
        "renewable_corn_oil_yield",
        "distillers_yield",
        "uhp_yield",
        "electricity_usage",
    }
    economics_overlay_source_support = EconomicsOverlaySourceSupport(
        EconomicsOverlaySourceSupportDeps(
            source_lines=source_lines,
            economics_market_rows=economics_market_rows,
            coefficient_templates=coefficient_templates,
            as_of_market_quarter=as_of_market_quarter,
            driver_source_display=_driver_source_display,
            driver_source_note=_driver_source_note,
            parse_driver_number=_parse_driver_number,
            convert_market_price_value=_convert_market_price_value,
            economics_market_region_tags=_economics_market_region_tags,
            quarter_label_short=_quarter_label_short,
        )
    )
    _best_line = economics_overlay_source_support.best_line
    _parse_market_input_value = economics_overlay_source_support.parse_market_input_value
    _driver_source_comment = economics_overlay_source_support.driver_source_comment
    _market_quality_rank = economics_overlay_source_support.market_quality_rank
    _pick_market_reference = economics_overlay_source_support.pick_market_reference
    _parse_quarter_label_text = economics_overlay_source_support.parse_quarter_label_text
    _overlay_coefficient_detail = economics_overlay_source_support.overlay_coefficient_detail
    _overlay_coefficient_basis_display = economics_overlay_source_support.overlay_coefficient_basis_display
    _overlay_coefficient_source_display = economics_overlay_source_support.overlay_coefficient_source_display
    _market_source_note = economics_overlay_source_support.market_source_note

    def _gpre_proxy_implied_frame_record(frame_key: str) -> Dict[str, Any]:
        if not (is_gpre_profile and gpre_commercial_setup_rows):
            return {}
        frame_map = (gpre_proxy_implied_results_bundle or {}).get("frames") or {}
        frame = frame_map.get(str(frame_key or "")) if isinstance(frame_map, dict) else {}
        return dict(frame) if isinstance(frame, dict) else {}

    row_map: Dict[Tuple[str, date], Dict[str, Any]] = {}
    _record_writer_substage("write_excel.drivers.render.economics_overlay.setup", overlay_setup_started)
    for rec in rows:
        dkey = str(rec.get("_driver_key") or "")
        qd = rec.get("Quarter")
        if dkey and isinstance(qd, date):
            row_map[(dkey, qd)] = rec
    bridge_bundle_map = _load_operating_driver_bridge_bundle_map(quarter_set)
    market_input_templates = list(getattr(company_profile, "economics_overlay_market_inputs", ()) or [])
    hidden_overlay_market_input_keys = {
        "distillers_grains_price",
        "uhp_price",
        "renewable_corn_oil_price",
        "soybean_oil_price_proxy",
        "corn_oil_premium_assumption",
        "implied_renewable_corn_oil_proxy_price",
    }
    hidden_overlay_process_keys = {
        "distillers_contribution",
        "uhp_contribution",
        "corn_oil_contribution",
        "coproduct_credit",
    }
    hedge_templates = list(getattr(company_profile, "economics_overlay_hedge_templates", ()) or [])
    bridge_templates = list(getattr(company_profile, "economics_overlay_bridge_rows", ()) or [])
    gpre_commercial_setup_rows = _gpre_commercial_setup_records_shared() if is_gpre_profile else []
    overlay_gpre_end_col = 21 if (is_gpre_profile and gpre_commercial_setup_rows) else 8
    if is_gpre_profile and gpre_commercial_setup_rows:
        overlay_display_quarters = [qd for qd in quarter_set if isinstance(qd, date) and qd >= date(2023, 3, 31)]
        if not overlay_display_quarters:
            overlay_display_quarters = quarter_set[-7:]
    else:
        overlay_display_quarters = quarter_set
    market_state_result = build_economics_overlay_market_state(
        EconomicsOverlayMarketStateDeps(
            is_gpre_profile=is_gpre_profile,
            ticker=ticker,
            ticker_roots=ticker_roots,
            economics_market_rows=economics_market_rows,
            row_map=row_map,
            overlay_display_quarters=overlay_display_quarters,
            overlay_market_as_of=overlay_market_as_of,
            gpre_commercial_setup_rows=gpre_commercial_setup_rows,
            cache_dir=cache_dir,
            company_profile=company_profile,
            state=state,
            overlay_coefficient_detail=_overlay_coefficient_detail,
            parse_quarter_label_text=_parse_quarter_label_text,
            quarter_label_short=_quarter_label_short,
            record_writer_substage=_record_writer_substage,
            build_gpre_plant_capacity_history=build_gpre_plant_capacity_history,
            load_or_download_gpre_corn_bids_snapshot=load_or_download_gpre_corn_bids_snapshot,
            fetch_gpre_corn_bids_snapshot=fetch_gpre_corn_bids_snapshot,
            build_gpre_official_proxy_snapshot=build_gpre_official_proxy_snapshot,
            build_gpre_official_proxy_history_series=build_gpre_official_proxy_history_series,
            build_prior_quarter_simple_crush_snapshot=build_prior_quarter_simple_crush_snapshot,
            build_current_qtd_simple_crush_snapshot=build_current_qtd_simple_crush_snapshot,
            build_simple_crush_history_series=build_simple_crush_history_series,
            build_next_quarter_thesis_snapshot=build_next_quarter_thesis_snapshot,
            market_input_fingerprint=market_input_fingerprint,
            data_root_from_sec_cache_path=data_root_from_sec_cache_path,
            build_gpre_basis_proxy_model=build_gpre_basis_proxy_model,
            build_gpre_overlay_proxy_preview_bundle=build_gpre_overlay_proxy_preview_bundle,
            resolve_gpre_quarter_open_snapshot=resolve_gpre_quarter_open_snapshot,
            market_build_gpre_proxy_implied_results_bundle=market_build_gpre_proxy_implied_results_bundle,
            persist_gpre_frozen_thesis_snapshot=persist_gpre_frozen_thesis_snapshot,
        )
    )
    gpre_ticker_root_local = market_state_result.gpre_ticker_root_local
    gpre_bids_snapshot = dict(market_state_result.gpre_bids_snapshot or {})
    gpre_plant_capacity_history = dict(market_state_result.gpre_plant_capacity_history or {})
    prior_q_market_snapshot = dict(market_state_result.prior_q_market_snapshot or {})
    current_qtd_market_snapshot = dict(market_state_result.current_qtd_market_snapshot or {})
    next_quarter_thesis_snapshot = dict(market_state_result.next_quarter_thesis_snapshot or {})
    simple_crush_history_rows = list(market_state_result.simple_crush_history_rows or [])
    gpre_basis_model_result = dict(market_state_result.gpre_basis_model_result or {})
    prior_market_status = market_state_result.prior_market_status
    current_market_status = market_state_result.current_market_status
    prior_market_available = market_state_result.prior_market_available
    current_market_available = market_state_result.current_market_available
    prior_market_display_quarter = market_state_result.prior_market_display_quarter
    current_market_display_quarter = market_state_result.current_market_display_quarter
    prior_market_display_quarter_txt = market_state_result.prior_market_display_quarter_txt
    current_market_display_quarter_txt = market_state_result.current_market_display_quarter_txt
    next_thesis_quarter_end = market_state_result.next_thesis_quarter_end
    next_thesis_quarter_txt = market_state_result.next_thesis_quarter_txt
    prior_process_status = market_state_result.prior_process_status
    current_process_status = market_state_result.current_process_status
    gpre_reported_margin_by_quarter = dict(market_state_result.gpre_reported_margin_by_quarter or {})
    gpre_underlying_margin_by_quarter = dict(market_state_result.gpre_underlying_margin_by_quarter or {})
    gpre_denominator_policy_by_quarter = dict(market_state_result.gpre_denominator_policy_by_quarter or {})
    gpre_reported_gallons_by_quarter = dict(market_state_result.gpre_reported_gallons_by_quarter or {})
    gpre_reported_gallons_sold_by_quarter = dict(market_state_result.gpre_reported_gallons_sold_by_quarter or {})
    gpre_reported_gallons_produced_by_quarter = dict(market_state_result.gpre_reported_gallons_produced_by_quarter or {})
    gpre_basis_quarter_map = dict(market_state_result.gpre_basis_quarter_map or {})
    gpre_basis_weights_latest = list(market_state_result.gpre_basis_weights_latest or [])
    gpre_official_market_rows = list(market_state_result.gpre_official_market_rows or [])
    gpre_official_market_summary = market_state_result.gpre_official_market_summary
    gpre_official_weighting_method = market_state_result.gpre_official_weighting_method
    gpre_official_ethanol_method = market_state_result.gpre_official_ethanol_method
    gpre_official_basis_method = market_state_result.gpre_official_basis_method
    gpre_official_gas_method = market_state_result.gpre_official_gas_method
    gpre_official_fallback_policy = market_state_result.gpre_official_fallback_policy
    gpre_overlay_preview_bundle = dict(market_state_result.gpre_overlay_preview_bundle or {})
    gpre_best_forward_preview_bundle = dict(market_state_result.gpre_best_forward_preview_bundle or {})
    gpre_proxy_implied_results_bundle = dict(market_state_result.gpre_proxy_implied_results_bundle or {})
    gpre_current_qtd_trend_tracking = dict(market_state_result.gpre_current_qtd_trend_tracking or {})
    quarter_open_market_snapshot = dict(market_state_result.quarter_open_market_snapshot or {})
    quarter_open_proxy_status = market_state_result.quarter_open_proxy_status
    quarter_open_provenance = market_state_result.quarter_open_provenance
    quarter_open_display_quarter = market_state_result.quarter_open_display_quarter
    quarter_open_display_quarter_txt = market_state_result.quarter_open_display_quarter_txt
    quarter_open_subheader_txt = market_state_result.quarter_open_subheader_txt
    chosen_preview_quality = market_state_result.chosen_preview_quality
    chosen_preview_mae = market_state_result.chosen_preview_mae
    chosen_preview_max_error = market_state_result.chosen_preview_max_error
    chosen_preview_top_miss = market_state_result.chosen_preview_top_miss
    chosen_preview_worst_phase = market_state_result.chosen_preview_worst_phase
    quarterly_df = market_state_result.quarterly_df

    def _overlay_market_date_text(value_in: Any) -> str:
        if isinstance(value_in, date):
            return value_in.isoformat()
        return "n/a"

    def _snapshot_market_meta(snapshot_in: Optional[Dict[str, Any]], key_in: str) -> Dict[str, Any]:
        meta_map = snapshot_in.get("market_meta") if isinstance(snapshot_in, dict) else {}
        meta = meta_map.get(str(key_in or "")) if isinstance(meta_map, dict) else {}
        return meta if isinstance(meta, dict) else {}

    def _current_qtd_summary_text() -> str:
        if is_gpre_profile and gpre_commercial_setup_rows:
            return (
                "Approximate market crush with natural gas cost, GPRE corn basis, and weighted to active capacity | "
                "GPRE crush proxy is a fitted model to approximate actual GPRE results."
            )
        if current_process_status != "ok":
            return (
                "Approximate market crush is the pre-hedge simple market/process proxy, excluding coproduct credits. "
                "Current QTD stays blank until overlapping current-quarter corn, ethanol and gas observations exist."
            )
        return "Approximate market crush is the pre-hedge simple market/process proxy, excluding coproduct credits"

    def _market_input_intro_text() -> str:
        if is_gpre_profile and gpre_commercial_setup_rows:
            return (
                "Prior quarter is latest reported quarter; Current QTD uses observed market data | "
                "Quarter-open outlook is current quarter at quarter-start pricing | "
                "Next quarter outlook is current futures prices for next quarter."
            )
        return (
            "Prior quarter and Current QTD use USDA market data where available, "
            "while Next quarter outlook uses local Chicago ethanol futures plus USDA futures for corn and natural gas."
        )

    def _first_gpre_ticker_root_local() -> Optional[Path]:
        return gpre_ticker_root_local

    def _market_futures_component_note(meta_in: Dict[str, Any], *, corn: bool = False) -> str:
        if not isinstance(meta_in, dict):
            return ""
        component_key = "futures_contract_components" if corn else "contract_components"
        components = [dict(item) for item in list(meta_in.get(component_key) or []) if isinstance(item, dict)]
        rule_txt = str(meta_in.get("futures_weighting_rule" if corn else "weighting_rule") or "").strip()
        source_files = [
            Path(str(item or "")).name
            for item in list(meta_in.get("source_files" if not corn else "futures_source_files") or [])
            if str(item or "").strip()
        ]
        if corn and not source_files:
            source_file_txt = str(meta_in.get("futures_source_file") or "").strip()
            source_files = [Path(item.strip()).name for item in source_file_txt.split(",") if item.strip()]
        if not source_files:
            source_file_txt = str(meta_in.get("source_file") or "").strip()
            source_files = [Path(item.strip()).name for item in source_file_txt.split(",") if item.strip()]
        if not components and not rule_txt and not source_files:
            return ""
        part_txts: List[str] = []
        for comp in components[:4]:
            symbol_txt = str(comp.get("symbol") or comp.get("contract_tenor") or "").strip().upper()
            obs_txt = _overlay_market_date_text(comp.get("observation_date"))
            price_num = pd.to_numeric(comp.get("price_value"), errors="coerce")
            weight_num = pd.to_numeric(comp.get("weight"), errors="coerce")
            price_txt = f"{float(price_num):.4f}".rstrip("0").rstrip(".") if pd.notna(price_num) else ""
            weight_txt = ""
            if pd.notna(weight_num) and abs(float(weight_num) - 1.0) > 1e-9:
                weight_txt = f", {float(weight_num):.0%}"
            segment = " ".join(item for item in (symbol_txt, obs_txt, price_txt) if item)
            if segment:
                part_txts.append(f"{segment}{weight_txt}")
        notes: List[str] = []
        if rule_txt:
            notes.append(f"Futures rule: {rule_txt}")
        if part_txts:
            notes.append("Contracts: " + "; ".join(part_txts))
        if source_files:
            notes.append("Files: " + ", ".join(source_files[:4]))
        fallback_reason = str(meta_in.get("fallback_reason") or "").strip()
        if fallback_reason:
            missing_txt = ", ".join(
                str(item or "").strip()
                for item in list(meta_in.get("fallback_missing_contract_tenors") or [])
                if str(item or "").strip()
            )
            notes.append(f"Fallback: {fallback_reason}{f' Missing: {missing_txt}' if missing_txt else ''}")
        return (" " + ". ".join(notes) + ".") if notes else ""

    def _market_component_symbol_summary(meta_in: Dict[str, Any], *, corn: bool = False) -> str:
        component_key = "futures_contract_components" if corn else "contract_components"
        components = [dict(item) for item in list((meta_in or {}).get(component_key) or []) if isinstance(item, dict)]
        symbols: List[str] = []
        for comp in components:
            symbol_txt = str(comp.get("symbol") or comp.get("contract_tenor") or "").strip().upper()
            if symbol_txt and symbol_txt not in symbols:
                symbols.append(symbol_txt)
        return "/".join(symbols)

    def _compact_corn_basis_source_label(label_in: Any) -> str:
        label_txt = str(label_in or "").strip()
        low = label_txt.lower()
        if "gpre" in low and "ams" in low:
            return "GPRE bids/AMS"
        if "gpre" in low:
            return "GPRE bids"
        if "ams" in low:
            return "AMS basis"
        return label_txt or "basis"

    def _market_override_for_frame(
        snapshot_in: Optional[Dict[str, Any]],
        *,
        key_in: str,
        period_label: str,
        period_quarter_txt: str,
    ) -> Optional[Dict[str, Any]]:
        current_market = snapshot_in.get("current_market") if isinstance(snapshot_in, dict) else {}
        current_market = current_market if isinstance(current_market, dict) else {}
        meta = _snapshot_market_meta(snapshot_in, key_in)
        value_out = current_market.get(str(key_in or ""))
        if is_gpre_profile and gpre_commercial_setup_rows and key_in == "corn_price":
            corn_basis_label = str(meta.get("official_corn_basis_source_label") or "").strip() or "weighted AMS basis proxy"
            corn_basis_provenance = str(meta.get("official_corn_basis_provenance") or "").strip()
            payload = (
                f"Delivered corn proxy (CBOT + {corn_basis_label})",
                (
                    f"front-month CBOT corn plus {corn_basis_label} aligned to the quarter frame, "
                    "then averaged across included observations."
                ),
                "Weeks included",
            )
        elif is_gpre_profile and gpre_commercial_setup_rows and key_in == "ethanol_price":
            payload = (
                "Weighted GPRE ethanol benchmark",
                "the footprint-weighted ethanol benchmark across the mapped GPRE regional/state ethanol series.",
                "Weeks included",
            )
        else:
            field_map = {
                "corn_price": (
                    "Nebraska daily cash average",
                    "Nebraska daily cash-market averages aligned to the quarter frame, then averaged across included observations.",
                    "Days included",
                ),
                "ethanol_price": (
                    "Nebraska weekly ethanol cash average",
                    "the weekly Nebraska ethanol cash average across included observations.",
                    "Weeks included",
                ),
                "natural_gas_price": (
                    "Front-month NYMEX Natural Gas",
                    "the front-month NYMEX natural-gas settlement across included observations.",
                    "Weeks included",
                ),
            }
            payload = field_map.get(str(key_in or ""))
        if payload is None:
            return None
        basis_txt, comment_prefix, count_label = payload
        as_of_val = meta.get("as_of")
        obs_count = int(meta.get("obs_count") or 0)
        as_of_txt = _overlay_market_date_text(as_of_val)
        available = value_out is not None and obs_count > 0
        if not available:
            explicit_message = str((snapshot_in or {}).get("message") or "").strip() if isinstance(snapshot_in, dict) else ""
            unavailable_txt = (
                f"{period_label} ({period_quarter_txt}) unavailable; no quarter-bucketed observations were available."
                if period_quarter_txt
                else f"{period_label} unavailable; no quarter-bucketed observations were available."
            )
            if explicit_message and str((snapshot_in or {}).get("status") or "").strip().lower() in {"no_snapshot", "no_data"}:
                unavailable_txt = explicit_message
            return {
                "value": None,
                "basis": f"{period_label}: unavailable",
                "comment": unavailable_txt,
                "available": False,
                "period_label": period_label,
                "period_quarter_txt": period_quarter_txt,
                "as_of_txt": "",
                "obs_count": 0,
                "count_label": count_label,
            }
        context_prefix = f"{period_label} ({period_quarter_txt})" if period_quarter_txt else period_label
        extra_comment = ""
        if is_gpre_profile and gpre_commercial_setup_rows and key_in == "corn_price":
            corn_basis_provenance = str(meta.get("official_corn_basis_provenance") or "").strip()
            if corn_basis_provenance:
                extra_comment = f" {corn_basis_provenance}"
        if is_gpre_profile and gpre_commercial_setup_rows and period_label in {"Quarter-open outlook", "Next quarter outlook"}:
            extra_comment = f"{extra_comment}{_market_futures_component_note(meta, corn=(key_in == 'corn_price'))}"
        source_summary = ""
        if is_gpre_profile and gpre_commercial_setup_rows:
            if period_label == "Quarter-open outlook":
                if key_in == "corn_price":
                    symbols_txt = _market_component_symbol_summary(meta, corn=True) or "local corn futures"
                    source_summary = f"Q-open: {symbols_txt} + {_compact_corn_basis_source_label(meta.get('official_corn_basis_source_label'))}"
                elif key_in == "natural_gas_price":
                    symbols_txt = _market_component_symbol_summary(meta) or "local gas futures"
                    source_summary = f"Q-open: {symbols_txt} local strip"
                elif key_in == "ethanol_price":
                    symbols_txt = _market_component_symbol_summary(meta) or "local ethanol futures"
                    source_summary = f"Q-open: {symbols_txt} local strip"
            elif period_label == "Current QTD":
                source_summary = "Current QTD: observed actuals"
        return {
            "value": value_out,
            "basis": f"{period_label}: {basis_txt}",
            "comment": f"{context_prefix} uses {comment_prefix} As of {as_of_txt}; {count_label.lower()}={obs_count}.{extra_comment}",
            "available": True,
            "period_label": period_label,
            "period_quarter_txt": period_quarter_txt,
            "as_of_txt": as_of_txt,
            "obs_count": obs_count,
            "count_label": count_label,
            "source_summary": source_summary,
        }

    def _prior_market_override(key_in: str) -> Optional[Dict[str, Any]]:
        return _market_override_for_frame(
            prior_q_market_snapshot,
            key_in=key_in,
            period_label="Prior quarter",
            period_quarter_txt=prior_market_display_quarter_txt,
        )

    def _current_market_override(key_in: str) -> Optional[Dict[str, Any]]:
        return _market_override_for_frame(
            current_qtd_market_snapshot,
            key_in=key_in,
            period_label="Current QTD",
            period_quarter_txt=current_market_display_quarter_txt,
        )

    def _quarter_open_market_override(key_in: str) -> Optional[Dict[str, Any]]:
        return _market_override_for_frame(
            quarter_open_market_snapshot,
            key_in=key_in,
            period_label="Quarter-open outlook",
            period_quarter_txt=quarter_open_display_quarter_txt,
        )

    def _thesis_market_override(key_in: str) -> Optional[Dict[str, Any]]:
        if key_in == "ethanol_price":
            ref = next_quarter_thesis_snapshot.get("ethanol") if isinstance(next_quarter_thesis_snapshot, dict) else None
            if not isinstance(ref, dict):
                return {
                    "value": None,
                    "manual": False,
                    "thesis_label": "Local Chicago ethanol futures unavailable",
                    "basis_suffix": "",
                    "comment": "Next-quarter ethanol thesis is unavailable because no local Chicago ethanol futures strip was resolved for the target quarter.",
                }
            price_val = pd.to_numeric(ref.get("price_value"), errors="coerce")
            contract_labels = [str(label or "").strip() for label in list(ref.get("contract_labels") or []) if str(label or "").strip()]
            contract_tenors = [str(label or "").strip() for label in list(ref.get("contract_tenors") or []) if str(label or "").strip()]
            strip_method = str(ref.get("strip_method") or "").strip() or "day_weighted"
            obs_txt = _overlay_market_date_text(ref.get("observation_date"))
            if pd.isna(price_val):
                missing_txt = ", ".join(str(item or "").strip() for item in list(ref.get("missing_contract_tenors") or []) if str(item or "").strip())
                return {
                    "value": None,
                    "manual": False,
                    "thesis_label": "Local Chicago ethanol futures unavailable",
                    "basis_suffix": "",
                    "comment": (
                        "Next-quarter ethanol thesis could not be built from the local Chicago ethanol futures strip."
                        + (f" Missing contract months: {missing_txt}." if missing_txt else "")
                    ),
                }
            label_txt = ", ".join(contract_labels) if contract_labels else ", ".join(contract_tenors)
            method_txt = "day-weighted quarterly strip" if strip_method == "day_weighted" else "simple three-contract average"
            component_note = _market_futures_component_note(
                {
                    "contract_components": list(ref.get("contract_components") or []),
                    "weighting_rule": str(ref.get("weighting_rule") or "").strip(),
                    "source_files": list(ref.get("source_files") or []),
                }
            )
            symbols_txt = "/".join(
                str((comp or {}).get("symbol") or (comp or {}).get("contract_tenor") or "").strip().upper()
                for comp in list(ref.get("contract_components") or [])
                if isinstance(comp, dict) and str((comp or {}).get("symbol") or (comp or {}).get("contract_tenor") or "").strip()
            )
            return {
                "value": float(price_val),
                "manual": False,
                "thesis_label": f"Local Chicago ethanol futures strip {label_txt}".strip(),
                "basis_suffix": method_txt,
                "comment": (
                    f"Next quarter outlook uses the local Chicago ethanol futures strip across {label_txt or 'the target-quarter contract months'} "
                    f"with a {method_txt}. Latest futures date used: {obs_txt}.{component_note}"
                ),
                "source_summary": f"Next: {symbols_txt or 'local ethanol futures'} local strip",
            }
        if key_in == "corn_price":
            ref = next_quarter_thesis_snapshot.get("corn") if isinstance(next_quarter_thesis_snapshot, dict) else None
        elif key_in == "natural_gas_price":
            ref = next_quarter_thesis_snapshot.get("natural_gas") if isinstance(next_quarter_thesis_snapshot, dict) else None
        else:
            ref = None
        if not isinstance(ref, dict):
            return None
        obs_txt = _overlay_market_date_text(ref.get("observation_date"))
        contract_label = str(ref.get("contract_label") or str(ref.get("contract_tenor") or "")).strip()
        price_val = pd.to_numeric(ref.get("price_value"), errors="coerce")
        if pd.isna(price_val):
            return None
        if is_gpre_profile and gpre_commercial_setup_rows and key_in == "corn_price":
            representative_basis = pd.to_numeric(ref.get("official_weighted_corn_basis_usd_per_bu"), errors="coerce")
            basis_label = str(ref.get("official_corn_basis_source_label") or "").strip() or "weighted AMS basis proxy"
            basis_provenance = str(ref.get("official_corn_basis_provenance") or "").strip()
            latest_basis_quarter = None
            if pd.isna(representative_basis):
                for qd_local, rec_local in sorted(gpre_basis_quarter_map.items()):
                    basis_num = pd.to_numeric((rec_local or {}).get("weighted_basis_plant_count_usd_per_bu"), errors="coerce")
                    if pd.notna(basis_num):
                        representative_basis = float(basis_num)
                        latest_basis_quarter = qd_local
                        basis_label = "weighted AMS basis proxy"
            thesis_value = float(price_val)
            thesis_label = f"CBOT Corn futures {contract_label}"
            basis_suffix = "futures-based approximation"
            if pd.notna(representative_basis):
                thesis_value = float(price_val) + float(representative_basis)
                thesis_label = f"CBOT Corn futures {contract_label} + {basis_label}"
                basis_suffix = basis_label
            rep_quarter_txt = _quarter_label_short(latest_basis_quarter) if isinstance(latest_basis_quarter, date) else ""
            basis_explainer = basis_provenance or (
                f"Representative weighted AMS basis {f'from {rep_quarter_txt}' if rep_quarter_txt else 'from the latest covered quarter'}."
            )
            component_note = _market_futures_component_note(
                {
                    "futures_contract_components": list(ref.get("contract_components") or []),
                    "futures_weighting_rule": str(ref.get("weighting_rule") or "").strip(),
                    "futures_source_file": str(ref.get("source_file") or "").strip(),
                },
                corn=True,
            )
            symbols_txt = "/".join(
                str((comp or {}).get("symbol") or (comp or {}).get("contract_tenor") or "").strip().upper()
                for comp in list(ref.get("contract_components") or [])
                if isinstance(comp, dict) and str((comp or {}).get("symbol") or (comp or {}).get("contract_tenor") or "").strip()
            )
            return {
                "value": thesis_value,
                "manual": False,
                "thesis_label": thesis_label,
                "basis_suffix": basis_suffix,
                "comment": (
                    f"Next quarter outlook uses the {contract_label} corn contract plus {basis_label} for the simple market row. "
                    f"{basis_explainer}{component_note} "
                    "Any additional timing or hedge effects are evaluated separately in the GPRE crush proxy."
                ),
                "source_summary": f"Next: {symbols_txt or contract_label or 'corn futures'} + {_compact_corn_basis_source_label(basis_label)}",
            }
        instrument_txt = "CBOT Corn futures" if key_in == "corn_price" else "NYMEX Natural Gas futures"
        component_note = _market_futures_component_note(
            {
                "contract_components": list(ref.get("contract_components") or []),
                "weighting_rule": str(ref.get("weighting_rule") or "").strip(),
                "source_file": str(ref.get("source_file") or "").strip(),
                "fallback_reason": str(ref.get("fallback_reason") or "").strip(),
                "fallback_missing_contract_tenors": list(ref.get("fallback_missing_contract_tenors") or []),
            }
        )
        symbols_txt = "/".join(
            str((comp or {}).get("symbol") or (comp or {}).get("contract_tenor") or "").strip().upper()
            for comp in list(ref.get("contract_components") or [])
            if isinstance(comp, dict) and str((comp or {}).get("symbol") or (comp or {}).get("contract_tenor") or "").strip()
        )
        return {
            "value": float(price_val),
            "manual": False,
            "thesis_label": f"{instrument_txt} {contract_label}",
            "basis_suffix": "futures-based approximation",
            "comment": (
                f"Next quarter outlook uses the {contract_label} contract because its contract-month midpoint is nearest the target-quarter midpoint. "
                f"Latest observation used: {obs_txt}.{component_note} Futures are an approximation and may differ from realized regional pricing."
            ),
            "source_summary": f"Next: {symbols_txt or contract_label or 'futures'} local strip",
        }

    def _market_frame_metadata_text(override_in: Optional[Dict[str, Any]]) -> str:
        if not isinstance(override_in, dict):
            return ""
        period_label = str(override_in.get("period_label") or "").strip()
        period_quarter_txt = str(override_in.get("period_quarter_txt") or "").strip()
        if not bool(override_in.get("available")):
            if period_label and period_quarter_txt:
                return f"{period_label} ({period_quarter_txt}) unavailable"
            return f"{period_label or 'Frame'} unavailable"
        as_of_txt = str(override_in.get("as_of_txt") or "").strip()
        obs_count = int(override_in.get("obs_count") or 0)
        count_label = str(override_in.get("count_label") or "Obs included").strip()
        period_prefix = f"{period_label} ({period_quarter_txt})" if period_quarter_txt else period_label
        return f"{period_prefix} | As of {as_of_txt} | {count_label}: {obs_count}"

    def _market_frame_status_text(override_in: Optional[Dict[str, Any]]) -> str:
        if not isinstance(override_in, dict):
            return "Current QTD unavailable"
        period_label = str(override_in.get("period_label") or "Current QTD").strip()
        period_quarter_txt = str(override_in.get("period_quarter_txt") or "").strip()
        period_prefix = f"{period_label} ({period_quarter_txt})" if period_quarter_txt else period_label
        if not bool(override_in.get("available")):
            return f"{period_prefix} unavailable"
        return f"{period_prefix} available"

    def _overlay_as_of_header_text(value_in: Any) -> str:
        if isinstance(value_in, date):
            return f"As of {value_in.isoformat()}"
        return "As of n/a"

    quarter_open_overlay_header_txt = _overlay_as_of_header_text(
        quarter_open_market_snapshot.get("as_of")
        if isinstance(quarter_open_market_snapshot, dict) and quarter_open_market_snapshot.get("as_of")
        else None
    )

    def _market_input_metadata_text(
        key_in: str,
        prior_override: Optional[Dict[str, Any]],
        quarter_open_override: Optional[Dict[str, Any]],
        current_override: Optional[Dict[str, Any]],
        thesis_override: Optional[Dict[str, Any]],
    ) -> str:
        parts: List[str] = []
        quarter_open_status_txt = _market_frame_status_text(quarter_open_override)
        current_status_txt = _market_frame_status_text(current_override)
        if quarter_open_status_txt:
            parts.append(quarter_open_status_txt)
        quarter_open_comment_txt = str(((quarter_open_override or {}).get("comment") or "")).strip()
        if quarter_open_comment_txt and (
            "No frozen prior-quarter thesis snapshot" in quarter_open_comment_txt
            or "Futures rule:" in quarter_open_comment_txt
            or "Fallback:" in quarter_open_comment_txt
        ):
            parts.append(quarter_open_comment_txt)
        if current_status_txt:
            parts.append(current_status_txt)
        thesis_label = str(((thesis_override or {}).get("thesis_label") or "")).strip()
        if thesis_label:
            parts.append(f"Thesis: {thesis_label}")
        thesis_comment_txt = str(((thesis_override or {}).get("comment") or "")).strip()
        if thesis_comment_txt and ("Futures rule:" in thesis_comment_txt or "Contracts:" in thesis_comment_txt):
            parts.append(thesis_comment_txt)
        return " | ".join(part for part in parts if part)

    def _market_input_basis_text(
        prior_override: Optional[Dict[str, Any]],
        quarter_open_override: Optional[Dict[str, Any]],
        current_override: Optional[Dict[str, Any]],
        thesis_override: Optional[Dict[str, Any]],
    ) -> str:
        basis_parts = []
        prior_basis_txt = str(((prior_override or {}).get("basis") or "")).strip()
        quarter_open_basis_txt = str(((quarter_open_override or {}).get("basis") or "")).strip()
        current_basis_txt = str(((current_override or {}).get("basis") or "")).strip()
        thesis_basis_suffix = str(((thesis_override or {}).get("basis_suffix") or "")).strip()
        if prior_basis_txt:
            basis_parts.append(prior_basis_txt)
        if quarter_open_basis_txt:
            basis_parts.append(quarter_open_basis_txt)
        if current_basis_txt:
            basis_parts.append(current_basis_txt)
        if thesis_basis_suffix:
            basis_parts.append(thesis_basis_suffix)
        return " | ".join(part for part in basis_parts if part)

    def _market_input_source_text(
        key_in: str,
        prior_override: Optional[Dict[str, Any]],
        quarter_open_override: Optional[Dict[str, Any]],
        current_override: Optional[Dict[str, Any]],
        thesis_override: Optional[Dict[str, Any]],
        source_txt: str,
    ) -> str:
        if not (is_gpre_profile and gpre_commercial_setup_rows):
            return str(source_txt or "").strip()
        def _status_text(override_in: Optional[Dict[str, Any]], *, keep_quarter: bool) -> str:
            if not isinstance(override_in, dict):
                return ""
            period_label = str(override_in.get("period_label") or "").strip()
            period_quarter_txt = str(override_in.get("period_quarter_txt") or "").strip()
            available = bool(override_in.get("available"))
            if period_label.lower().startswith("quarter-open"):
                label_txt = f"Quarter-open outlook ({period_quarter_txt})" if (keep_quarter and period_quarter_txt) else "Quarter-open outlook"
            elif period_label.lower().startswith("current qtd"):
                label_txt = "Current QTD"
            else:
                label_txt = period_label or "Current QTD"
            return f"{label_txt}: {'available' if available else 'unavailable'}"

        def _quarter_open_short_text(override_in: Optional[Dict[str, Any]]) -> str:
            if not isinstance(override_in, dict):
                return ""
            available = bool(override_in.get("available"))
            if not available:
                return _status_text(override_in, keep_quarter=True)
            if quarter_open_provenance == "manual_local_snapshot":
                return "Quarter-open outlook uses local manual snapshot."
            if quarter_open_provenance == "frozen_snapshot":
                return "Quarter-open outlook uses frozen prior-quarter snapshot."
            return _status_text(override_in, keep_quarter=True)

        def _thesis_short_text(thesis_override_in: Optional[Dict[str, Any]], key_local: str) -> str:
            source_summary = str(((thesis_override_in or {}).get("source_summary") or "")).strip()
            if source_summary:
                return source_summary
            thesis_label = str(((thesis_override_in or {}).get("thesis_label") or "")).strip()
            basis_suffix = str(((thesis_override_in or {}).get("basis_suffix") or "")).strip()
            thesis_value = pd.to_numeric((thesis_override_in or {}).get("value"), errors="coerce")
            combined = " ".join(part for part in (thesis_label, basis_suffix) if part).lower()
            if key_local == "corn_price":
                if "actual gpre plant-bid basis" in combined and "ams fallback" in combined:
                    return "Next quarter outlook uses live bids + AMS fallback."
                if "weighted ams" in combined:
                    return "Next quarter outlook uses weighted AMS proxy."
            if key_local == "natural_gas_price" and "nymex" in combined:
                return "Next quarter outlook uses NYMEX futures."
            if key_local == "ethanol_price":
                if "local chicago ethanol futures" in combined and not pd.isna(thesis_value):
                    return "Next quarter outlook uses local Chicago ethanol futures strip."
                    return "Next quarter outlook ethanol unavailable."
            return f"Next quarter outlook: {thesis_label}" if thesis_label else ""

        parts: List[str] = []
        quarter_open_txt = str(((quarter_open_override or {}).get("source_summary") or "")).strip() or _quarter_open_short_text(quarter_open_override)
        current_txt = str(((current_override or {}).get("source_summary") or "")).strip() or _status_text(current_override, keep_quarter=False)
        thesis_txt = _thesis_short_text(thesis_override, key_in)
        if quarter_open_txt:
            parts.append(quarter_open_txt)
        if current_txt:
            parts.append(current_txt)
        if thesis_txt:
            parts.append(thesis_txt)
        if not parts:
            return str(source_txt or "").strip()
        return " | ".join(part for part in parts if part)

    current_overlay_model_key = str(gpre_basis_model_result.get("gpre_proxy_model_key") or "").strip()
    best_forward_overlay_model_key = str(gpre_basis_model_result.get("best_forward_lens_model_key") or "").strip()
    overlay_model_key_to_pred_col = (
        dict(gpre_basis_model_result.get("model_key_to_pred_col") or {})
        if isinstance(gpre_basis_model_result, dict)
        else {}
    )
    overlay_leaderboard_df = (
        gpre_basis_model_result.get("leaderboard_df")
        if isinstance(gpre_basis_model_result.get("leaderboard_df"), pd.DataFrame)
        else pd.DataFrame()
    )

    def _overlay_model_leaderboard_row(model_key_in: Any) -> Dict[str, Any]:
        key_txt = str(model_key_in or "").strip()
        if not key_txt or not isinstance(overlay_leaderboard_df, pd.DataFrame) or overlay_leaderboard_df.empty:
            return {}
        sub = overlay_leaderboard_df[overlay_leaderboard_df["model_key"].astype(str) == key_txt].copy()
        return sub.iloc[0].to_dict() if not sub.empty else {}

    def _overlay_preview_bundle_for_model(model_key_in: Any) -> Dict[str, Any]:
        key_txt = str(model_key_in or "").strip()
        if not key_txt or key_txt == current_overlay_model_key:
            return dict(gpre_overlay_preview_bundle or {})
        if key_txt == best_forward_overlay_model_key:
            return dict(gpre_best_forward_preview_bundle or {})
        return {}

    def _gpre_preview_frame_value(frame_group: str, frame_key: str) -> Optional[float]:
        frame = (((gpre_overlay_preview_bundle or {}).get(frame_group) or {}).get(frame_key) or {})
        value_num = pd.to_numeric((frame or {}).get("value"), errors="coerce")
        if pd.isna(value_num):
            return None
        return float(value_num)

    def _gpre_preview_frame_status(frame_group: str, frame_key: str) -> str:
        frame = (((gpre_overlay_preview_bundle or {}).get(frame_group) or {}).get(frame_key) or {})
        return str((frame or {}).get("status") or "no_data")

    def _gpre_preview_frame_note(frame_group: str, frame_key: str) -> str:
        def _fallback_note(frame_key_in: str) -> str:
            phase_map = {
                "quarter_open": "quarter_open",
                "current_qtd": "current",
                "next_quarter_thesis": "next",
            }
            preview_story = market_gpre_phase_preview_story(
                current_overlay_model_key,
                phase=phase_map.get(str(frame_key_in or ""), str(frame_key_in or "")),
            )
            note_txt = str((preview_story or {}).get("live_preview_note") or "").strip()
            if note_txt:
                return note_txt
            if str(frame_key_in or "") == "quarter_open":
                return "Quarter-open fitted value for the chosen model."
            return ""
        frame = (((gpre_overlay_preview_bundle or {}).get(frame_group) or {}).get(frame_key) or {})
        frame_note = str((frame or {}).get("live_preview_note") or "").strip()
        return frame_note or _fallback_note(str(frame_key or ""))

    def _gpre_model_preview_frame_value(model_key: str, frame_key: str) -> Optional[float]:
        frame = (((_overlay_preview_bundle_for_model(model_key) or {}).get("gpre_proxy_frames") or {}).get(frame_key) or {})
        value_num = pd.to_numeric((frame or {}).get("value"), errors="coerce")
        if pd.isna(value_num):
            return None
        return float(value_num)

    def _gpre_model_preview_frame_note(model_key: str, frame_key: str) -> str:
        frame = (((_overlay_preview_bundle_for_model(model_key) or {}).get("gpre_proxy_frames") or {}).get(frame_key) or {})
        frame_note = str((frame or {}).get("live_preview_note") or "").strip()
        if frame_note:
            return frame_note
        phase_map = {
            "quarter_open": "quarter_open",
            "current_qtd": "current",
            "next_quarter_thesis": "next",
        }
        preview_story = market_gpre_phase_preview_story(
            str(model_key or current_overlay_model_key),
            phase=phase_map.get(str(frame_key or ""), str(frame_key or "")),
        )
        note_txt = str((preview_story or {}).get("live_preview_note") or "").strip()
        if note_txt:
            return note_txt
        if str(frame_key or "") == "quarter_open":
            return "Quarter-open fitted value for the chosen model."
        return ""

    def _same_quarter_last_year(qd_in: Any) -> Optional[date]:
        if not isinstance(qd_in, date):
            return None
        try:
            return date(qd_in.year - 1, qd_in.month, qd_in.day)
        except Exception:
            return None

    def _quarter_end_for_date_local(dt_in: Any) -> Optional[date]:
        if not isinstance(dt_in, date):
            return None
        quarter_end_month = ((int(dt_in.month) - 1) // 3 + 1) * 3
        quarter_end_day = 31 if quarter_end_month in {3, 12} else 30
        return date(dt_in.year, quarter_end_month, quarter_end_day)

    def _historical_proxy_value(qd_in: Any, *, fitted: bool, model_key: str = "") -> Optional[float]:
        if not isinstance(qd_in, date):
            return None
        rec = dict(gpre_basis_quarter_map.get(qd_in) or {})
        if not rec:
            return None
        col_name = "official_simple_proxy_usd_per_gal"
        if fitted:
            resolved_model_key = str(model_key or current_overlay_model_key).strip()
            if resolved_model_key and resolved_model_key != current_overlay_model_key:
                pred_col = str(overlay_model_key_to_pred_col.get(resolved_model_key) or "").strip()
                col_name = pred_col or "gpre_proxy_official_usd_per_gal"
            else:
                col_name = "gpre_proxy_official_usd_per_gal"
        value_num = pd.to_numeric(rec.get(col_name), errors="coerce")
        if pd.isna(value_num):
            return None
        return float(value_num)

    def _format_proxy_comp_value(value_in: Any) -> str:
        value_num = pd.to_numeric(value_in, errors="coerce")
        if pd.isna(value_num):
            return ""
        return f"{float(value_num):.3f}"

    def _format_yoy_comp_text(current_value: Any, ly_value: Any, *, unavailable_label: str = "Unavailable") -> str:
        current_num = pd.to_numeric(current_value, errors="coerce")
        ly_num = pd.to_numeric(ly_value, errors="coerce")
        if pd.isna(current_num) and pd.isna(ly_num):
            return unavailable_label
        if pd.isna(current_num):
            return f"n/a vs {float(ly_num):.3f}"
        if pd.isna(ly_num):
            return f"{float(current_num):.3f} vs n/a"
        delta = float(current_num) - float(ly_num)
        return f"{float(current_num):.3f} vs {float(ly_num):.3f} ({delta:+.3f})"

    def _ordered_market_input_templates() -> List[Any]:
        priority = {"corn_price": 0, "natural_gas_price": 1, "ethanol_price": 2}
        ranked: List[Tuple[int, int, Any]] = []
        for idx, tpl in enumerate(market_input_templates):
            key = str(getattr(tpl, "key", "") or "").strip()
            ranked.append((priority.get(key, 100 + idx), idx, tpl))
        return [tpl for _, _, tpl in sorted(ranked, key=lambda item: (item[0], item[1]))]

    def _write_overlay_intro(
        row_num: int,
        text_in: str,
        *,
        end_col: int = 8,
        spacer_after: int = 1,
        row_height: Optional[float] = None,
    ) -> int:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=end_col)
        cell = ws.cell(row=row_num, column=1, value=text_in)
        cell.fill = intro_fill
        cell.font = Font(size=font_size, color=dark_text_color)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for cc in range(1, end_col + 1):
            ws.cell(row=row_num, column=cc).fill = intro_fill
            ws.cell(row=row_num, column=cc).border = row_border
        ws.row_dimensions[row_num].height = float(row_height if row_height is not None else 34)
        for offs in range(1, int(spacer_after) + 1):
            ws.row_dimensions[row_num + offs].height = overlay_spacer_row_height
        return row_num + 1 + int(spacer_after)

    def _write_overlay_subheader_row(
        row_num: int,
        *,
        prior_txt: str,
        quarter_open_txt: str,
        current_txt: str,
        thesis_txt: str,
        note_start_col: int,
        note_end_col: int,
        row_height: Optional[float] = None,
    ) -> int:
        spans = [
            (1, 1, ""),
            (2, 3, prior_txt),
            (4, 5, quarter_open_txt),
            (6, 7, current_txt),
            (8, 9, thesis_txt),
            (10, 10, ""),
            (11, note_start_col - 1 if note_start_col > 11 else 10, ""),
            (note_start_col, note_end_col, ""),
        ]
        for start_col, end_col, label in spans:
            if end_col < start_col:
                continue
            if end_col > start_col:
                ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
            cell = ws.cell(row=row_num, column=start_col, value=label)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center" if start_col in {2, 4, 6, 8} else "left", vertical="center", wrap_text=True)
            cell.border = thin_border
            for cc in range(start_col, end_col + 1):
                ws.cell(row=row_num, column=cc).fill = header_fill
                ws.cell(row=row_num, column=cc).border = thin_border
                ws.cell(row=row_num, column=cc).alignment = Alignment(horizontal="center" if start_col in {2, 4, 6, 8} else "left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_num].height = float(row_height if row_height is not None else 21.0)
        return row_num + 1

    def _center_header_span(row_num: int, start_col: int, end_col: int) -> None:
        for cc in range(start_col, end_col + 1):
            ws.cell(row=row_num, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_year_band(row_num: int, label: str, *, end_col: int = 8, row_height: Optional[float] = None) -> int:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=end_col)
        cell = ws.cell(row=row_num, column=1, value=label)
        cell.font = year_band_font
        cell.fill = year_band_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for cc in range(1, end_col + 1):
            ws.cell(row=row_num, column=cc).fill = year_band_fill
            ws.cell(row=row_num, column=cc).border = row_border
        ws.row_dimensions[row_num].height = float(row_height if row_height is not None else 18)
        return row_num + 1

    bridge_separator_rows: List[int] = []
    gpre_bridge_panel_rows: Dict[str, int] = {}
    market_section_bar_row: Optional[int] = None

    def _apply_bridge_to_reported_section(row_num: int) -> int:
        nonlocal bridge_separator_rows, gpre_bridge_panel_rows, gpre_reported_margin_by_quarter, gpre_denominator_policy_by_quarter
        bridge_result = write_gpre_economics_overlay_bridge_to_reported_section(
            GpreEconomicsOverlayBridgeDeps(
                ws=ws,
                row_idx=row_num,
                is_gpre_profile=is_gpre_profile,
                gpre_commercial_setup_rows=gpre_commercial_setup_rows,
                overlay_display_quarters=overlay_display_quarters,
                overlay_gpre_end_col=overlay_gpre_end_col,
                row_map=row_map,
                bridge_bundle_map=bridge_bundle_map,
                derivative_bridge_by_quarter=derivative_bridge_by_quarter,
                bridge_templates=bridge_templates,
                market_input_templates=market_input_templates,
                gpre_basis_quarter_map=gpre_basis_quarter_map,
                current_overlay_model_key=current_overlay_model_key,
                best_forward_overlay_model_key=best_forward_overlay_model_key,
                overlay_model_key_to_pred_col=overlay_model_key_to_pred_col,
                gpre_bridge_panel_rows=gpre_bridge_panel_rows,
                gpre_reported_margin_by_quarter=gpre_reported_margin_by_quarter,
                gpre_denominator_policy_by_quarter=gpre_denominator_policy_by_quarter,
                overlay_section_row_height=overlay_section_row_height,
                overlay_intro_row_height=overlay_intro_row_height,
                overlay_header_row_height=overlay_header_row_height,
                overlay_support_row_height=overlay_support_row_height,
                header_fill=header_fill,
                thin_border=thin_border,
                bold_font=bold_font,
                body_font=body_font,
                zebra_fill_light=zebra_fill_light,
                write_section_bar=_write_section_bar,
                write_overlay_intro=_write_overlay_intro,
                add_comment=_add_comment,
                overlay_coefficient_detail=_overlay_coefficient_detail,
                pick_market_reference=_pick_market_reference,
                overlay_model_leaderboard_row=_overlay_model_leaderboard_row,
                overlay_model_label=_overlay_model_label,
                driver_source_comment=_driver_source_comment,
                driver_source_note=_driver_source_note,
            )
        )
        bridge_separator_rows = list(bridge_result.bridge_separator_rows)
        gpre_bridge_panel_rows = dict(bridge_result.gpre_bridge_panel_rows)
        gpre_reported_margin_by_quarter = dict(bridge_result.gpre_reported_margin_by_quarter)
        gpre_denominator_policy_by_quarter = dict(bridge_result.gpre_denominator_policy_by_quarter)
        return bridge_result.row_idx

    title_end_col = overlay_gpre_end_col if (is_gpre_profile and gpre_commercial_setup_rows) else max(8, 1 + len(overlay_display_quarters))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_end_col)
    tcell = ws.cell(row=1, column=1, value="Economics Overlay")
    tcell.font = title_font
    tcell.fill = title_fill
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(1, title_end_col + 1):
        ws.cell(row=1, column=cc).fill = title_fill
    ws.row_dimensions[1].height = 24

    if is_gpre_profile and gpre_commercial_setup_rows:
        ws.column_dimensions["A"].width = _px_to_width(315.0)
        for letter in tuple("BCDEFGHIJKLMNOPQRSTU"):
            ws.column_dimensions[letter].width = _px_to_width(102.0)
    else:
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 42
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 44

    coeff_rows: Dict[str, int] = {}
    market_rows: Dict[str, int] = {}
    row_idx = 3
    if is_gpre_profile and gpre_commercial_setup_rows:
        row_idx = 4
        commercial_result = write_gpre_economics_overlay_commercial_sections(
            GpreEconomicsOverlayCommercialDeps(
                ws=ws,
                is_gpre_profile=is_gpre_profile,
                row_idx=row_idx,
                gpre_commercial_setup_rows=gpre_commercial_setup_rows,
                derivative_bridge_by_quarter=derivative_bridge_by_quarter,
                overlay_gpre_end_col=overlay_gpre_end_col,
                analysis_theme=analysis_theme,
                body_font=body_font,
                bold_font=bold_font,
                horizon_font=horizon_font,
                setup_font=setup_font,
                row_border=row_border,
                thin_border=thin_border,
                quarter_separator_side=quarter_separator_side,
                overlay_commentary_section_row_height=overlay_commentary_section_row_height,
                overlay_commentary_header_row_height=overlay_commentary_header_row_height,
                overlay_commentary_year_band_row_height=overlay_commentary_year_band_row_height,
                overlay_commercial_section_row_height=overlay_commercial_section_row_height,
                overlay_commercial_header_row_height=overlay_commercial_header_row_height,
                overlay_commercial_year_band_row_height=overlay_commercial_year_band_row_height,
                overlay_commercial_row_max_height=overlay_commercial_row_max_height,
                overlay_support_row_height=overlay_support_row_height,
                add_comment=_add_comment,
                ensure_terminal_period=_ensure_terminal_period,
                estimate_wrapped_row_height=_estimate_wrapped_row_height,
                normalize_text=glx_normalize_text,
                overlay_driver_source_priority=_overlay_driver_source_priority,
                record_writer_substage=_record_writer_substage,
                write_header_row=_write_header_row,
                write_section_bar=_write_section_bar,
                write_year_band=_write_year_band,
            )
        )
        row_idx = commercial_result.row_idx
        overlay_bridge_started = time.perf_counter()
        row_idx = _apply_bridge_to_reported_section(row_idx)
        ws.row_dimensions[82].height = 24.0
        _record_writer_substage("write_excel.drivers.render.economics_overlay.bridge_to_reported", overlay_bridge_started)
        row_idx += 1
    input_rows_result = write_gpre_economics_overlay_input_rows(
        GpreEconomicsOverlayInputRowsDeps(
            ws=ws,
            row_idx=row_idx,
            is_gpre_profile=is_gpre_profile,
            gpre_commercial_setup_rows=gpre_commercial_setup_rows,
            overlay_gpre_end_col=overlay_gpre_end_col,
            overlay_display_quarters=overlay_display_quarters,
            coefficient_templates=coefficient_templates,
            gpre_official_market_rows=gpre_official_market_rows,
            gpre_basis_quarter_map=gpre_basis_quarter_map,
            gpre_official_market_summary=gpre_official_market_summary,
            gpre_official_weighting_method=gpre_official_weighting_method,
            gpre_official_ethanol_method=gpre_official_ethanol_method,
            gpre_official_basis_method=gpre_official_basis_method,
            gpre_official_gas_method=gpre_official_gas_method,
            gpre_official_fallback_policy=gpre_official_fallback_policy,
            hidden_overlay_coefficient_keys=hidden_overlay_coefficient_keys,
            hidden_overlay_market_input_keys=hidden_overlay_market_input_keys,
            prior_market_display_quarter_txt=prior_market_display_quarter_txt,
            quarter_open_overlay_header_txt=quarter_open_overlay_header_txt,
            current_qtd_market_snapshot=current_qtd_market_snapshot,
            next_thesis_quarter_txt=next_thesis_quarter_txt,
            prior_q_market_snapshot=prior_q_market_snapshot,
            quarter_open_market_snapshot=quarter_open_market_snapshot,
            next_quarter_thesis_snapshot=next_quarter_thesis_snapshot,
            current_overlay_model_key=current_overlay_model_key,
            overlay_section_row_height=overlay_section_row_height,
            overlay_intro_row_height=overlay_intro_row_height,
            overlay_header_row_height=overlay_header_row_height,
            overlay_support_row_height=overlay_support_row_height,
            thin_border=thin_border,
            body_font=body_font,
            input_fill=input_fill,
            input_font=input_font,
            section_fill=section_fill,
            align_left_center_wrap=align_left_center_wrap,
            align_center=align_center,
            align_left_center=align_left_center,
            align_center_wrap=align_center_wrap,
            align_left_top_wrap=align_left_top_wrap,
            intro_fill=intro_fill,
            zebra_fill_light=zebra_fill_light,
            zebra_fill_dark=zebra_fill_dark,
            bold_font=bold_font,
            header_fill=header_fill,
            font_size=font_size,
            dark_text_color=dark_text_color,
            write_section_bar=_write_section_bar,
            write_overlay_intro=_write_overlay_intro,
            write_header_row=_write_header_row,
            center_header_span=_center_header_span,
            write_overlay_subheader_row=_write_overlay_subheader_row,
            overlay_coefficient_detail=_overlay_coefficient_detail,
            overlay_coefficient_basis_display=_overlay_coefficient_basis_display,
            overlay_coefficient_source_display=_overlay_coefficient_source_display,
            add_comment=_add_comment,
            record_writer_substage=_record_writer_substage,
            market_input_intro_text=_market_input_intro_text,
            ordered_market_input_templates=_ordered_market_input_templates,
            pick_market_reference=_pick_market_reference,
            market_source_note=_market_source_note,
            driver_source_note=_driver_source_note,
            prior_market_override=_prior_market_override,
            quarter_open_market_override=_quarter_open_market_override,
            current_market_override=_current_market_override,
            thesis_market_override=_thesis_market_override,
            market_input_source_text=_market_input_source_text,
            overlay_as_of_header_text=_overlay_as_of_header_text,
            overlay_preview_bundle_for_model=_overlay_preview_bundle_for_model,
            snapshot_market_meta=_snapshot_market_meta,
            market_gpre_phase_preview_story=market_gpre_phase_preview_story,
        )
    )
    row_idx = input_rows_result.row_idx
    coeff_rows = dict(input_rows_result.coeff_rows)
    market_rows = dict(input_rows_result.market_rows)
    coeff_ref = dict(input_rows_result.coeff_ref)
    prior_ref = dict(input_rows_result.prior_ref)
    quarter_open_ref = dict(input_rows_result.quarter_open_ref)
    current_ref = dict(input_rows_result.current_ref)
    thesis_ref = dict(input_rows_result.thesis_ref)
    market_section_bar_row = input_rows_result.market_section_bar_row
    _write_gpre_approx_market_crush_build_up_section = input_rows_result.write_gpre_approx_market_crush_build_up_section
    _gpre_fitted_live_formula = input_rows_result.gpre_fitted_live_formula
    _gpre_formula_note = input_rows_result.gpre_formula_note
    _gpre_model_live_formula = input_rows_result.gpre_model_live_formula
    _gpre_model_formula_note = input_rows_result.gpre_model_formula_note
    def _write_gpre_basis_proxy_sandbox(
        target_ws: Any,
        start_row: int,
        model_result: Dict[str, Any],
    ) -> Dict[str, Any]:
        basis_proxy_sandbox_deps = BasisProxySandboxWriterDeps(
            extract_operating_driver_rows_for_template=_extract_operating_driver_rows_for_template,
            load_operating_driver_source_records_by_quarter=_load_operating_driver_source_records_by_quarter,
            load_operating_driver_template_index=_load_operating_driver_template_index,
            market_quality_rank=_market_quality_rank,
            operating_driver_quarters=_operating_driver_quarters,
            overlay_coefficient_detail=_overlay_coefficient_detail,
            overlay_market_date_text=_overlay_market_date_text,
            quarter_label_short=_quarter_label_short,
            write_gpre_approx_market_crush_build_up_section=_write_gpre_approx_market_crush_build_up_section,
            align_center=align_center,
            align_center_wrap=align_center_wrap,
            analysis_theme=analysis_theme,
            as_of_market_quarter=as_of_market_quarter,
            body_font=body_font,
            bold_font=bold_font,
            border_color=border_color,
            coeff_rows=coeff_rows,
            current_market_display_quarter=current_market_display_quarter,
            dark_text_color=dark_text_color,
            economics_market_rows=economics_market_rows,
            font_size=font_size,
            gpre_commercial_setup_rows=gpre_commercial_setup_rows,
            gpre_plant_capacity_history=gpre_plant_capacity_history,
            gpre_proxy_implied_results_bundle=gpre_proxy_implied_results_bundle,
            gpre_reported_gallons_by_quarter=gpre_reported_gallons_by_quarter,
            gpre_ticker_root_local=gpre_ticker_root_local,
            header_fill=header_fill,
            header_size=header_size,
            intro_fill=intro_fill,
            is_gpre_profile=is_gpre_profile,
            market_input_templates_by_key=market_input_templates_by_key,
            market_rows=market_rows,
            muted_text_color=muted_text_color,
            next_thesis_quarter_end=next_thesis_quarter_end,
            prior_market_display_quarter=prior_market_display_quarter,
            quarter_open_display_quarter=quarter_open_display_quarter,
            row_map=row_map,
            section_fill=section_fill,
            thin_border=thin_border,
            zebra_fill_dark=zebra_fill_dark,
            zebra_fill_light=zebra_fill_light,
        )
        return write_basis_proxy_sandbox_sheet(
            basis_proxy_sandbox_deps,
            target_ws=target_ws,
            start_row=start_row,
            model_result=model_result,
        )

    _record_writer_substage("write_excel.drivers.render.economics_overlay.market_inputs", input_rows_result.market_inputs_started)
    overlay_support_result = write_gpre_basis_proxy_overlay_support(
        GpreOverlaySupportInputs(
            wb=wb,
            ws=ws,
            row_idx=row_idx,
            is_gpre_profile=is_gpre_profile,
            has_gpre_commercial_setup=bool(gpre_commercial_setup_rows),
            model_result=gpre_basis_model_result,
            proxy_implied_results_bundle=gpre_proxy_implied_results_bundle,
            bridge_panel_rows=gpre_bridge_panel_rows,
            ticker_root=_first_gpre_ticker_root_local(),
            current_ref=current_ref,
            thesis_ref=thesis_ref,
            prior_quarter_header_text=prior_market_display_quarter_txt,
            quarter_open_header_text=quarter_open_overlay_header_txt,
            current_qtd_header_text=_overlay_as_of_header_text(current_qtd_market_snapshot.get("as_of") if isinstance(current_qtd_market_snapshot, dict) else None),
            next_quarter_header_text=next_thesis_quarter_txt,
            title_fill=title_fill,
            title_font=title_font,
            header_fill=header_fill,
            bold_font=bold_font,
            body_font=body_font,
            thin_border=thin_border,
            align_center=align_center,
            align_center_wrap=align_center_wrap,
            align_left_center_wrap=align_left_center_wrap,
            zebra_fill_light=zebra_fill_light,
            sandbox_writer=_write_gpre_basis_proxy_sandbox,
            add_comment=_add_comment,
            gpre_fitted_live_formula=_gpre_fitted_live_formula,
            gpre_formula_note=_gpre_formula_note,
            gpre_preview_frame_value=_gpre_preview_frame_value,
            gpre_preview_frame_note=_gpre_preview_frame_note,
            gpre_model_live_formula=_gpre_model_live_formula,
            gpre_model_formula_note=_gpre_model_formula_note,
            gpre_model_preview_frame_value=_gpre_model_preview_frame_value,
            gpre_model_preview_frame_note=_gpre_model_preview_frame_note,
            record_writer_substage=_record_writer_substage,
        )
    )
    gpre_basis_sandbox_layout = dict(overlay_support_result.sandbox_layout or {})
    sandbox_process_margin_refs = (
        ((gpre_basis_sandbox_layout.get("approx_market_crush_build_up") or {}).get("process_margin_refs"))
        if isinstance(gpre_basis_sandbox_layout, dict)
        else {}
    ) or {}
    proxy_comp_end_row = int(overlay_support_result.proxy_comp_end_row or (row_idx - 1))
    proxy_comp_title_row = int(overlay_support_result.proxy_comp_title_row or 0)
    proxy_comp_header_row = int(overlay_support_result.proxy_comp_header_row or 0)
    official_proxy_comp_row = int(overlay_support_result.official_proxy_comp_row or 0)
    fitted_proxy_comp_row = int(overlay_support_result.fitted_proxy_comp_row or 0)
    best_forward_proxy_comp_row = int(overlay_support_result.best_forward_proxy_comp_row or 0)

    write_gpre_derivative_crush_tests_side_effect(
        GpreEconomicsOverlayDerivativeSideEffectDeps(
            runtime_sheet_owned=derivative_crush_tests_owned,
            derivative_oci_bridge_df=derivative_oci_bridge_df,
            derivative_oci_exposure_df=derivative_oci_exposure_df,
            operating_driver_history_rows=operating_driver_history_rows,
            gpre_basis_model_result=gpre_basis_model_result,
            info_log=info_log,
            build_derivative_crush_tests=build_derivative_crush_tests,
            write_derivative_crush_tests_sheet=_write_derivative_crush_tests_sheet,
        )
    )

    quarter_compare_result = write_gpre_overlay_quarter_comparisons(
        GpreOverlayQuarterComparisonDeps(
            ws=ws,
            is_gpre_profile=is_gpre_profile,
            has_gpre_commercial_setup=bool(gpre_commercial_setup_rows),
            proxy_comp_end_row=proxy_comp_end_row,
            proxy_comp_title_row=proxy_comp_title_row,
            proxy_comp_header_row=proxy_comp_header_row,
            official_proxy_comp_row=official_proxy_comp_row,
            fitted_proxy_comp_row=fitted_proxy_comp_row,
            best_forward_proxy_comp_row=best_forward_proxy_comp_row,
            prior_market_display_quarter=prior_market_display_quarter,
            quarter_open_display_quarter=quarter_open_display_quarter,
            current_market_display_quarter=current_market_display_quarter,
            next_thesis_quarter_end=next_thesis_quarter_end,
            current_overlay_model_key=current_overlay_model_key,
            best_forward_overlay_model_key=best_forward_overlay_model_key,
            title_fill=analysis_theme["title_fill"],
            title_font=analysis_theme["title_font"],
            header_fill=header_fill,
            body_font=body_font,
            bold_font=bold_font,
            thin_border=thin_border,
            zebra_fill_light=zebra_fill_light,
            same_quarter_last_year=_same_quarter_last_year,
            historical_proxy_value=_historical_proxy_value,
            gpre_preview_frame_value=_gpre_preview_frame_value,
            gpre_model_preview_frame_value=_gpre_model_preview_frame_value,
            format_yoy_comp_text=_format_yoy_comp_text,
            record_writer_substage=_record_writer_substage,
        )
    )
    row_idx = quarter_compare_result.row_idx

    # Keep overlay charts on one visual grid and cap quarterly/coprod windows so
    # new quarters roll in without turning the delivered workbook into a horizontal dump.
    overlay_chart_width = 34.0
    overlay_chart_height = 16.0
    overlay_chart_row_span = 24
    overlay_quarter_chart_max_points = 15
    chart_result = write_economics_overlay_charts(
        EconomicsOverlayChartWriterDeps(
            ws=ws,
            row_idx=row_idx,
            is_gpre_profile=is_gpre_profile,
            gpre_commercial_setup_rows=gpre_commercial_setup_rows,
            simple_crush_history_rows=simple_crush_history_rows,
            proxy_comp_end_row=proxy_comp_end_row,
            official_proxy_comp_row=official_proxy_comp_row,
            fitted_proxy_comp_row=fitted_proxy_comp_row,
            best_forward_proxy_comp_row=best_forward_proxy_comp_row,
            next_quarter_thesis_snapshot=next_quarter_thesis_snapshot,
            sandbox_process_margin_refs=sandbox_process_margin_refs,
            thesis_ref=thesis_ref,
            prior_market_display_quarter=prior_market_display_quarter,
            quarter_open_display_quarter=quarter_open_display_quarter,
            current_market_display_quarter=current_market_display_quarter,
            next_thesis_quarter_end=next_thesis_quarter_end,
            quarterly_df=locals().get("quarterly_df"),
            overlay_model_key_to_pred_col=overlay_model_key_to_pred_col,
            current_overlay_model_key=current_overlay_model_key,
            best_forward_overlay_model_key=best_forward_overlay_model_key,
            title_fill=title_fill,
            title_font=title_font,
            thin_border=thin_border,
            align_center=align_center,
            add_comment=_add_comment,
            gpre_preview_frame_value=_gpre_preview_frame_value,
            gpre_model_preview_frame_value=_gpre_model_preview_frame_value,
            historical_proxy_value=_historical_proxy_value,
            apply_chart_text_categories=_apply_chart_text_categories,
            record_writer_substage=_record_writer_substage,
            chart_width=overlay_chart_width,
            chart_height=overlay_chart_height,
            chart_row_span=overlay_chart_row_span,
            max_chart_points=overlay_quarter_chart_max_points,
        )
    )
    row_idx = chart_result.row_idx

    current_qtd_result = write_gpre_economics_overlay_current_qtd_section(
        GpreEconomicsOverlayCurrentQtdDeps(
            ws=ws,
            row_idx=row_idx,
            is_gpre_profile=is_gpre_profile,
            has_gpre_commercial_setup=bool(gpre_commercial_setup_rows),
            gpre_current_qtd_trend_tracking=gpre_current_qtd_trend_tracking,
            title_fill=title_fill,
            title_font=title_font,
            header_fill=header_fill,
            body_font=body_font,
            bold_font=bold_font,
            thin_border=thin_border,
            zebra_fill_light=zebra_fill_light,
            zebra_fill_dark=zebra_fill_dark,
            intro_fill=intro_fill,
            align_center=align_center,
            parse_snapshot_date_like=_gpre_parse_snapshot_date_like,
            record_writer_substage=_record_writer_substage,
        )
    )
    row_idx = current_qtd_result.row_idx
    gpre_overlay_coproduct_start_row = current_qtd_result.coproduct_start_row

    coproduct_result = write_gpre_economics_overlay_coproduct_section(
        GpreEconomicsOverlayCoproductDeps(
            wb=wb,
            ws=ws,
            row_idx=row_idx,
            is_gpre_profile=is_gpre_profile,
            gpre_commercial_setup_rows=gpre_commercial_setup_rows,
            gpre_basis_sandbox_layout=gpre_basis_sandbox_layout,
            gpre_overlay_coproduct_start_row=gpre_overlay_coproduct_start_row,
            market_rows=market_rows,
            coeff_rows=coeff_rows,
            current_qtd_market_snapshot=current_qtd_market_snapshot,
            prior_market_display_quarter=prior_market_display_quarter,
            quarter_open_display_quarter=quarter_open_display_quarter,
            current_market_display_quarter=current_market_display_quarter,
            next_thesis_quarter_end=next_thesis_quarter_end,
            prior_market_display_quarter_txt=prior_market_display_quarter_txt,
            quarter_open_overlay_header_txt=quarter_open_overlay_header_txt,
            next_thesis_quarter_txt=next_thesis_quarter_txt,
            overlay_chart_width=overlay_chart_width,
            overlay_chart_height=overlay_chart_height,
            overlay_chart_row_span=overlay_chart_row_span,
            overlay_quarter_chart_max_points=overlay_quarter_chart_max_points,
            overlay_header_row_height=overlay_header_row_height,
            title_fill=title_fill,
            title_font=title_font,
            header_fill=header_fill,
            body_font=body_font,
            bold_font=bold_font,
            thin_border=thin_border,
            zebra_fill_light=zebra_fill_light,
            zebra_fill_dark=zebra_fill_dark,
            intro_fill=intro_fill,
            section_fill=section_fill,
            align_center=align_center,
            overlay_as_of_header_text=_overlay_as_of_header_text,
            overlay_coefficient_detail=_overlay_coefficient_detail,
            parse_quarter_label_text=_parse_quarter_label_text,
            quarter_label_short=_quarter_label_short,
            apply_chart_text_categories=_apply_chart_text_categories,
            write_overlay_subheader_row=_write_overlay_subheader_row,
            record_writer_substage=_record_writer_substage,
        )
    )
    row_idx = coproduct_result.row_idx

    overlay_final_formatting_started = time.perf_counter()
    if is_gpre_profile and gpre_commercial_setup_rows:
        for separator_row in bridge_separator_rows:
            ws.row_dimensions[separator_row].height = 12.0
        electricity_usage_row = coeff_rows.get("electricity_usage")
        if not (isinstance(electricity_usage_row, int) and electricity_usage_row > 0):
            electricity_usage_row = next(
                (
                    rr
                    for rr in range(1, int(ws.max_row or 0) + 1)
                    if str(ws.cell(row=rr, column=1).value or "").strip().lower() == "electricity usage"
                ),
                None,
            )
        if isinstance(electricity_usage_row, int) and electricity_usage_row > 0:
            ws.row_dimensions[electricity_usage_row].height = 33.0
        if isinstance(market_section_bar_row, int) and market_section_bar_row > 0:
            ws.row_dimensions[market_section_bar_row].height = 30.0
        for fixed_row, fixed_height in ((82, 24.0), (122, 18.0)):
            ws.row_dimensions[fixed_row].height = fixed_height
        for fixed_row, fixed_height in ((86, 22.5), (101, 22.5), (106, 21.0)):
            ws.row_dimensions[fixed_row].height = fixed_height

    if not (is_gpre_profile and gpre_commercial_setup_rows):
        note_row = row_idx
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
        ws.cell(row=note_row, column=1, value="Approximate pre-hedge, pre-bridge process proxy. Compare it first to underlying/process economics, not reported EBITDA. Use Bridge to reported to reconcile hedge, policy, accounting and non-ethanol effects.")
        ws.cell(row=note_row, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(row=note_row, column=1).border = thin_border
        ws.row_dimensions[note_row].height = 30
        row_idx += 2
    else:
        pass
        row_idx += 1
    _record_writer_substage("write_excel.drivers.render.economics_overlay.final_formatting", overlay_final_formatting_started)

    if not (is_gpre_profile and gpre_commercial_setup_rows):
        row_idx = _write_section_bar(row_idx, "Hedge / position overlay", end_col=8)
        row_idx = _write_header_row(row_idx, ["Commodity", "Exposure unit", "Open position / disclosed hedge", "Hedge type", "Direction", "As of", "Confidence", "Comment / source"])
        hedge_terms = ("hedg", "risk management", "futures", "swap", "option", "contract", "lock in")
        for tpl in hedge_templates:
            label = str(getattr(tpl, "label", "") or "")
            aliases = tuple(getattr(tpl, "aliases", ()) or (label,))
            hit = _best_line(
                aliases,
                extra_terms=hedge_terms,
                exclude_terms=("forward-looking statements", "forward looking statements", "table of contents", "business overview"),
                preferred_sources=("10-K", "10-Q", "earnings_release", "presentation", "transcript"),
            )
            line_txt = str((hit or {}).get("line") or "")
            low = line_txt.lower()
            source_doc_txt = str((hit or {}).get("record", {}).get("source_doc") or "").lower()
            frag_pen = float(_text_fragment_penalty(line_txt) or 0.0)
            qty_pat = re.search(r"([0-9]{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(bushels|gallons|lbs|pounds|tons|mmbtu)", line_txt, re.I)
            looks_like_ocr_garble = bool(
                re.search(r"(?:\b[A-Za-z]\b\s*){8,}", line_txt)
                or re.search(r"[A-Za-z]\s+[A-Za-z]\s+[A-Za-z]\s+[A-Za-z]\s+[A-Za-z]", line_txt)
            )
            if (
                "forward-looking" in low
                or "forward looking" in low
                or looks_like_ocr_garble
                or (frag_pen > 1.25 and qty_pat is None)
                or ("ocr" in source_doc_txt and qty_pat is None and not qn_is_complete_signal_text(line_txt))
                or (not qn_is_complete_signal_text(line_txt) and qty_pat is None and len(line_txt.split()) < 8)
            ):
                hit = None
                line_txt = ""
                low = ""
                source_doc_txt = ""
                frag_pen = 0.0
                qty_pat = None
            hedge_type = "futures" if "futures" in low else "swap" if "swap" in low else "option" if "option" in low else "forward" if "forward contract" in low else "contract" if "contract" in low else "risk management" if "risk management" in low else ""
            direction = "protective" if any(tok in low for tok in ("lock in", "supports", "protect")) else "sold/short" if any(tok in low for tok in ("sold", "sale")) else "bought/long" if any(tok in low for tok in ("purchased", "bought", "buy")) else ""
            open_pos = f"{qty_pat.group(1)} {str(qty_pat.group(2) or '').replace('pounds','lbs')}" if qty_pat else _truncate_driver_text(line_txt, 72)
            confidence = "high" if hedge_type and re.search(r"\b\d", line_txt) else "medium" if hedge_type else ""
            as_of_txt = hit.get("quarter").isoformat() if hit and isinstance(hit.get("quarter"), date) else ""
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=str(getattr(tpl, "exposure_unit", "") or ""))
            ws.cell(row=row_idx, column=3, value=open_pos if hit else "")
            ws.cell(row=row_idx, column=4, value=hedge_type)
            ws.cell(row=row_idx, column=5, value=direction)
            ws.cell(row=row_idx, column=6, value=as_of_txt)
            ws.cell(row=row_idx, column=7, value=confidence)
            ws.cell(row=row_idx, column=8, value=_truncate_driver_text(line_txt, 110) if hit else "")
            for cc in range(1, 9):
                ws.cell(row=row_idx, column=cc).border = thin_border
                ws.cell(row=row_idx, column=cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc in {1, 3, 8})
            if hit:
                _add_comment(f"H{row_idx}", _driver_source_note(hit.get("record", {}).get("source_doc"), hit.get("line")))
            ws.row_dimensions[row_idx].height = _estimate_wrapped_row_height(ws.cell(row=row_idx, column=8).value, float(ws.column_dimensions["H"].width or 44), 18, 12, min_lines=1, max_lines=4) if hit else 18
        row_idx += 1

        row_idx += 1
        row_idx = _apply_bridge_to_reported_section(row_idx)
