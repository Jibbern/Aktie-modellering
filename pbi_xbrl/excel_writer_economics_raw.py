"""economics_market_raw audit sheet writer extracted from excel_writer_context."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Callable, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ECONOMICS_MARKET_RAW_HEADERS = [
    "observation_date",
    "quarter",
    "aggregation_level",
    "source_file",
    "source_type",
    "market_family",
    "series_key",
    "instrument",
    "region",
    "contract_tenor",
    "price_value",
    "unit",
    "parsed_text",
    "quality",
]

ECONOMICS_MARKET_RAW_COLUMN_WIDTHS = {
    "A": 14,
    "B": 12,
    "C": 16,
    "D": 28,
    "E": 18,
    "F": 18,
    "G": 24,
    "H": 24,
    "I": 18,
    "J": 16,
    "K": 12,
    "L": 12,
    "M": 44,
    "N": 12,
}

ECONOMICS_MARKET_RAW_LARGE_FAST_PATH_THRESHOLD = 20000


@dataclass(frozen=True)
class EconomicsMarketRawWriterDeps:
    wb: Workbook
    header_size: float
    safe_cell: Callable[[Any], Any]
    estimate_wrapped_row_height: Callable[..., float]


def write_economics_market_raw_sheet(deps: EconomicsMarketRawWriterDeps, rows: List[Dict[str, Any]]) -> None:
    wb = deps.wb
    header_size = deps.header_size
    _safe_cell = deps.safe_cell
    _estimate_wrapped_row_height = deps.estimate_wrapped_row_height

    ws = wb.create_sheet("economics_market_raw")
    local_header_fill = PatternFill("solid", fgColor="F2F2F2")
    local_thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    headers = ECONOMICS_MARKET_RAW_HEADERS
    if not rows:
        ws["A1"] = "No economics market data available."
        return
    ws.append(headers)
    for cc, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=cc, value=header)
        cell.font = Font(bold=True, size=header_size)
        cell.fill = local_header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = local_thin_border
    col_widths = ECONOMICS_MARKET_RAW_COLUMN_WIDTHS
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width
    large_raw_sheet_fast_path = len(rows) > ECONOMICS_MARKET_RAW_LARGE_FAST_PATH_THRESHOLD
    if large_raw_sheet_fast_path:
        # This audit sheet can exceed 80k rows for GPRE market history. The
        # workbook-facing quality comes from preserving the data and headers; per-cell
        # borders/alignment on every raw row dominate write time without adding much
        # audit value.
        for rec in rows:
            out_row: List[Any] = []
            for header in headers:
                value = rec.get(header)
                if isinstance(value, str):
                    value = ILLEGAL_CHARACTERS_RE.sub("", value)
                elif value is not None:
                    try:
                        value = _safe_cell(value)
                    except Exception:
                        pass
                out_row.append(value)
            ws.append(out_row)
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=1).number_format = "yyyy-mm-dd"
            ws.cell(row=row_idx, column=2).number_format = "yyyy-mm-dd"
            ws.cell(row=row_idx, column=11).number_format = "#,##0.000"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:N{ws.max_row}"
        ws.sheet_format.defaultRowHeight = 18
        ws.sheet_view.zoomScale = 110
        return
    for row_idx, rec in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = rec.get(header)
            if isinstance(value, str):
                value = ILLEGAL_CHARACTERS_RE.sub("", value)
            elif value is not None:
                try:
                    value = _safe_cell(value)
                except Exception:
                    pass
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = local_thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top" if header == "parsed_text" else "center", wrap_text=header == "parsed_text")
        ws.cell(row=row_idx, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=row_idx, column=2).number_format = "yyyy-mm-dd"
        px_cell = ws.cell(row=row_idx, column=11)
        if pd.notna(pd.to_numeric(rec.get("price_value"), errors="coerce")):
            px_cell.number_format = "#,##0.000"
        note_txt = str(rec.get("parsed_text") or "").strip()
        # Most market-export provenance strings fit on one line. Avoid the
        # wrapped-height estimator for those tens of thousands of rows; it is
        # reserved for genuinely long text where row height affects readability.
        if note_txt and len(note_txt) > int(float(col_widths["M"]) * 1.4):
            ws.row_dimensions[row_idx].height = _estimate_wrapped_row_height(
                note_txt,
                float(col_widths["M"]),
                18,
                12,
                min_lines=1,
                max_lines=4,
            )
        else:
            ws.row_dimensions[row_idx].height = 18
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{ws.max_row}"
    ws.sheet_format.defaultRowHeight = 18
    ws.sheet_view.zoomScale = 110
