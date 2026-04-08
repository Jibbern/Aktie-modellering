"""Dedicated writer surface for the Hidden_Value_Flags sheet."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Callable

import pandas as pd
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


@dataclass(frozen=True)
class HiddenValueFlagsSheetInputs:
    wb: Any
    sheet_name: str
    flags_df: pd.DataFrame
    font_size: int
    header_size: int
    safe_cell: Callable[[Any], Any]


def _autowidth(ws: Any, n_cols: int) -> None:
    for idx in range(1, n_cols + 1):
        width = max(14, ws.column_dimensions[get_column_letter(idx)].width or 14)
        if idx == 1:
            width = max(width, 26)
        ws.column_dimensions[get_column_letter(idx)].width = width


def _try_add_table(ws: Any, sheet_name: str, headers: list[Any]) -> None:
    try:
        if len(headers) == len(set(headers)) and all(isinstance(h, str) for h in headers):
            ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
            table = Table(displayName=sheet_name.replace(" ", "").replace("-", ""), ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(table)
    except Exception:
        pass


def write_hidden_value_flags_sheet(inputs: HiddenValueFlagsSheetInputs) -> None:
    ws = inputs.wb.create_sheet(inputs.sheet_name)
    df = inputs.flags_df
    if df is None or df.empty:
        ws["A1"] = "No signals."
        return

    headers = list(df.columns)
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append([None if pd.isna(row[col]) else inputs.safe_cell(row[col]) for col in headers])

    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, size=inputs.header_size)
        cell.alignment = Alignment(vertical="center")
    ws.sheet_format.defaultRowHeight = 18
    ws.sheet_view.zoomScale = 110
    _autowidth(ws, len(headers))

    # Hidden_Value_Flags has a stable visible contract that Valuation formulas and
    # readback tests depend on, so widths and formatting stay surface-local here.
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["C"].width = 30
    if "K" in ws.column_dimensions:
        ws.column_dimensions["K"].width = max(34, min(42, ws.column_dimensions["K"].width or 34))

    col_map = {header: idx + 1 for idx, header in enumerate(headers)}
    for col_name in ["evidence_1", "evidence_2", "evidence_3"]:
        col_idx = col_map.get(col_name)
        if not col_idx:
            continue
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(34, min(38, ws.column_dimensions[letter].width or 34))
        for rr in range(2, ws.max_row + 1):
            ws[f"{letter}{rr}"].alignment = Alignment(wrap_text=True, vertical="top")

    score_idx = col_map.get("score")
    score_range = None
    if score_idx:
        letter = get_column_letter(score_idx)
        score_range = f"{letter}2:{letter}{ws.max_row}"
        ws.conditional_formatting.add(
            score_range,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["70"],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )

    if score_idx and score_range:
        row_range = f"A2:{get_column_letter(len(headers))}{ws.max_row}"
        ws.conditional_formatting.add(
            row_range,
            FormulaRule(formula=["AND($B2<>\"\",$D2>=70)"], fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        ws.conditional_formatting.add(
            row_range,
            FormulaRule(formula=["AND($B2<>\"\",$D2>=40,$D2<70)"], fill=PatternFill("solid", fgColor="FFEB9C")),
        )
        ws.conditional_formatting.add(
            row_range,
            FormulaRule(formula=["AND($B2<>\"\",$D2<40)"], fill=PatternFill("solid", fgColor="FFC7CE")),
        )
        ws.conditional_formatting.add(
            score_range,
            CellIsRule(
                operator="between",
                formula=["40", "69.999"],
                fill=PatternFill("solid", fgColor="FFEB9C"),
            ),
        )
        ws.conditional_formatting.add(
            score_range,
            CellIsRule(
                operator="lessThan",
                formula=["40"],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )
    elif {"Flag", "Title", "Status", "Why it failed", "Key blocker"}.issubset(set(headers)):
        for col_name, width in {"A": 10, "B": 28, "C": 22, "D": 56, "E": 28}.items():
            ws.column_dimensions[col_name].width = width
        for rr in range(2, ws.max_row + 1):
            for cc in range(1, 6):
                ws.cell(row=rr, column=cc).alignment = Alignment(wrap_text=True, vertical="top")
    elif {"Flag", "Status", "Why it failed", "Key blocker"}.issubset(set(headers)):
        for col_name, width in {"A": 12, "B": 24, "C": 56, "D": 28}.items():
            ws.column_dimensions[col_name].width = width
        for rr in range(2, ws.max_row + 1):
            for cc in range(1, 5):
                ws.cell(row=rr, column=cc).alignment = Alignment(wrap_text=True, vertical="top")

    _try_add_table(ws, inputs.sheet_name, headers)
