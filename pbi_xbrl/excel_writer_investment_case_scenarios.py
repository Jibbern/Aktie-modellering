"""Investment Case scenario bridge, tax, and driver-assumption support/renderers."""
from __future__ import annotations

import math
import re
from dataclasses import dataclass
from typing import Any, Dict, List, MutableMapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class InvestmentCaseScenarioRenderDeps:
    runtime: MutableMapping[str, Any]


SCENARIO_DRIVER_REVENUE_VOLUME = "revenue_volume"
SCENARIO_DRIVER_MARGIN_EBITDA = "margin_ebitda"
SCENARIO_DRIVER_CASH_FLOW_CAPEX = "cash_flow_capex"
SCENARIO_DRIVER_CAPITAL_STRUCTURE_INTEREST = "capital_structure_interest"
SCENARIO_DRIVER_SHARE_COUNT_BUYBACK = "share_count_buyback"
SCENARIO_DRIVER_TAX_CREDIT_SUBSIDY = "tax_credit_subsidy"
SCENARIO_DRIVER_MANUAL_INCREMENTAL = "manual_incremental"

SCENARIO_TAX_TAXABLE = "taxable"
SCENARIO_TAX_NON_TAXABLE = "non_taxable"
SCENARIO_TAX_NON_TAXABLE_CREDIT = "non_taxable_credit"
SCENARIO_TAX_CASH_ONLY = "cash_only"
SCENARIO_TAX_NO_EPS_IMPACT = "no_eps_impact"
SCENARIO_TAX_UNKNOWN_MANUAL_REQUIRED = "unknown_manual_required"
SCENARIO_TAX_DIRECT_EPS = "direct_eps"


@dataclass(frozen=True)
class _ScenarioDriverBridgeSpec:
    """Reusable Investment_Case scenario bridge row definition.

    Manual input rows show total values.  Bridge rows translate those totals into
    incremental impacts only when the baseline and driver classification are
    clean enough to avoid double-counting.
    """

    label: str
    driver_type: str
    baseline: Any
    active: Any
    read: str
    explicit_incremental: bool = False
    reverse_incremental: bool = False
    ebitda_impact: str = "auto"
    fcf_impact: str = "auto"
    eps_impact: str = "auto"
    subsidy_basis: str = ""
    tax_treatment: str = SCENARIO_TAX_UNKNOWN_MANUAL_REQUIRED
    tax_source_basis: str = ""
    eps_impact_rule: str = ""
    audit_notes: str = ""


@dataclass
class _SegmentScenarioInputSpec:
    """Visible Segment Scenario Inputs row plus audit metadata."""

    label: str
    category_type: str
    baseline_revenue_m: Optional[float] = None
    revenue_basis: str = ""
    margin_conversion: Any = None
    margin_basis: str = ""
    feeds_bridge: bool = False
    source_note: str = ""
    view_basis: str = ""
    revenue_change_ref: str = ""
    revenue_impact_ref: str = ""
    ebitda_impact_ref: str = ""
    feeds_bridge_ref: str = ""


def _segment_scenario_revenue_m(value: Any, unit: Any = "") -> Optional[float]:
    val = pd.to_numeric(value, errors="coerce")
    if pd.isna(val):
        return None
    out = float(val)
    unit_txt = str(unit or "").strip().lower()
    if out and (abs(out) > 10000.0 or unit_txt in {"$", "usd"}):
        out /= 1_000_000.0
    return out if math.isfinite(out) else None


def _segment_scenario_view_basis(label: Any, category_type: Any = "") -> str:
    text = f"{label or ''} {category_type or ''}".strip().lower()
    if "brand" in text:
        return "Brand"
    if "geography" in text or "stores" in text or "emea" in text or "apac" in text or "americas" in text:
        return "Geography"
    return ""


def _segment_scenario_margin_value(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith("="):
            return stripped
        if re.match(r"^\$?[A-Z]{1,3}\$?\d+$", stripped):
            return f'=IF({stripped}="","",{stripped})'
    margin_num = pd.to_numeric(value, errors="coerce")
    if pd.isna(margin_num):
        return None
    margin = float(margin_num)
    if abs(margin) > 1.5:
        margin /= 100.0
    return margin if math.isfinite(margin) else None


def _segment_scenario_note_for_view(note: str, view_basis: str) -> str:
    txt = str(note or "").strip()
    low = txt.lower()
    if "separate revenue cut" in low or "not summed" in low:
        if view_basis == "Brand":
            return "Summed if Brand selected"
        if view_basis == "Geography":
            return "Summed if Geography selected"
    return txt


def _segment_scenario_specs_from_records(
    records: Sequence[Dict[str, Any]],
    *,
    default_margin_proxy: Any = None,
    default_margin_basis: str = "Company operating margin proxy",
) -> List[_SegmentScenarioInputSpec]:
    specs: List[_SegmentScenarioInputSpec] = []
    for rec in records or []:
        label = str(rec.get("metric") or rec.get("segment") or rec.get("label") or "").strip()
        if not label:
            continue
        category_type = str(rec.get("segment_type") or rec.get("type") or "").strip() or "Segment / category"
        margin = _segment_scenario_margin_value(rec.get("operating_margin", rec.get("margin_conversion")))
        baseline_m = _segment_scenario_revenue_m(rec.get("value"), rec.get("unit"))
        margin_basis = str(rec.get("margin_basis") or "").strip()
        view_basis = str(rec.get("view_basis") or "").strip() or _segment_scenario_view_basis(label, category_type)
        used_default_margin = False
        if margin is None and baseline_m is not None and default_margin_proxy not in (None, ""):
            margin = _segment_scenario_margin_value(default_margin_proxy)
            margin_basis = margin_basis or default_margin_basis
            used_default_margin = margin is not None
        feeds_txt = str(rec.get("feeds_bridge") or "").strip().lower()
        feeds_bridge = (feeds_txt in {"yes", "true", "1"} or used_default_margin) and baseline_m is not None and margin is not None
        note = str(rec.get("source_note") or rec.get("notes") or rec.get("note") or "").strip()
        if used_default_margin and (
            not note
            or "missing segment margin" in note.lower()
            or "company operating margin proxy" in note.lower()
            or "separate revenue cut" in note.lower()
            or note.lower().startswith("summed if ")
        ):
            note = default_margin_basis if default_margin_basis else "Company operating margin proxy"
        elif not note:
            if baseline_m is None:
                note = "Missing segment revenue"
            elif margin is None:
                note = "Missing segment margin"
            elif "company operating margin proxy" in margin_basis.lower():
                note = margin_basis
            elif not feeds_bridge:
                note = "Informational only"
        note = _segment_scenario_note_for_view(note, view_basis)
        specs.append(
            _SegmentScenarioInputSpec(
                label=label,
                category_type=category_type,
                baseline_revenue_m=baseline_m,
                revenue_basis=str(rec.get("revenue_basis") or rec.get("source") or "").strip(),
                margin_conversion=margin,
                margin_basis=margin_basis,
                feeds_bridge=feeds_bridge,
                source_note=note,
                view_basis=view_basis,
            )
        )
    return specs


def write_scenario_driver_assumptions_sheet(
    deps: InvestmentCaseScenarioRenderDeps,
    *,
    ticker: Any,
    segment_specs: Sequence[_SegmentScenarioInputSpec] = (),
    enabled: bool = True,
    disabled_note: str = "",
) -> None:
    runtime = deps.runtime
    wb = runtime["wb"]
    sheet_name = "Scenario_Driver_Assumptions"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ticker_txt = str(ticker or "").strip().upper()
    preferred_after = None
    tax_sheet = "Scenario_Bridge_Tax_Treatment"
    data_sheet = f"{ticker_txt}_Investment_Case_Data" if ticker_txt else ""
    case_sheet = f"{ticker_txt}_Investment_Case" if ticker_txt else ""
    for candidate in (tax_sheet, data_sheet, case_sheet):
        if candidate in wb.sheetnames:
            preferred_after = wb.sheetnames.index(candidate) + 1
            break
    ws = wb.create_sheet(sheet_name, index=preferred_after if preferred_after is not None else None)
    headers = [
        "Ticker",
        "Section",
        "Segment / category",
        "Type",
        "Revenue basis",
        "Baseline revenue",
        "Revenue % change",
        "Revenue impact",
        "Margin basis",
        "Operating margin",
        "EBITDA impact",
        "Feeds bridge?",
        "Source / note",
    ]
    header_fill = PatternFill("solid", fgColor="EAF3F8")
    thin = Border(
        left=Side(style="thin", color="D9E2EA"),
        right=Side(style="thin", color="D9E2EA"),
        top=Side(style="thin", color="D9E2EA"),
        bottom=Side(style="thin", color="D9E2EA"),
    )
    for cc, label in enumerate(headers, start=1):
        cell = ws.cell(1, cc, label)
        cell.fill = header_fill
        cell.font = Font(bold=True, size=11, color="1F2933")
        cell.border = thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    rows: List[List[Any]] = []
    if enabled:
        for spec in segment_specs:
            rows.append(
                [
                    ticker_txt,
                    "Segment Scenario Inputs",
                    spec.label,
                    spec.category_type,
                    spec.revenue_basis,
                    spec.baseline_revenue_m if spec.baseline_revenue_m is not None else "",
                    spec.revenue_change_ref,
                    spec.revenue_impact_ref,
                    spec.margin_basis,
                    spec.margin_conversion if spec.margin_conversion is not None else "",
                    spec.ebitda_impact_ref,
                    spec.feeds_bridge_ref or ("Yes" if spec.feeds_bridge else "No"),
                    spec.source_note,
                ]
            )
    else:
        rows.append(
            [
                ticker_txt,
                "Segment Scenario Inputs",
                "Not enabled",
                "Disabled",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "No",
                disabled_note or "Segment scenario disabled for this ticker profile.",
            ]
        )
    for rr, values in enumerate(rows, start=2):
        for cc, value in enumerate(values, start=1):
            cell = ws.cell(rr, cc, value)
            cell.border = thin
            cell.font = Font(size=10, color="1F2933")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if cc == 6:
                cell.number_format = "#,##0.0"
            if cc in {7, 10}:
                cell.number_format = "0.0%"
            if cc in {8, 11}:
                cell.number_format = "#,##0.0"
        ws.row_dimensions[rr].height = 24
    widths = {
        1: 10,
        2: 24,
        3: 34,
        4: 24,
        5: 30,
        6: 18,
        7: 18,
        8: 18,
        9: 28,
        10: 18,
        11: 18,
        12: 14,
        13: 42,
    }
    for cc, width in widths.items():
        ws.column_dimensions[get_column_letter(cc)].width = width
    ws.freeze_panes = "A2"


def _scenario_bridge_active_value_formula(ref: str) -> str:
    return f'=IF({ref}="","",{ref})'


def _excel_percent_value_expr(ref: str) -> str:
    """Return an Excel expression that accepts either 7% or 7 for percent inputs."""
    return f"IF({ref}>1,{ref}/100,{ref})"


def _excel_manual_percent_active_formula(row: int) -> str:
    def _candidate(ref: str) -> str:
        return _excel_percent_value_expr(ref)

    return (
        f'=IF(F{row}<>"",{_candidate(f"F{row}")},'
        f'IF(C{row}<>"",{_candidate(f"C{row}")},'
        f'IF(B{row}<>"",{_candidate(f"B{row}")},'
        f'IF(D{row}<>"",{_candidate(f"D{row}")},'
        f'IF(E{row}<>"",{_candidate(f"E{row}")},"")))))'
    )


def _excel_visible_value_range_formula(row: int) -> str:
    """Locale-safe text range for visible per-share scenario outputs."""
    return (
        f'=IFERROR(IF(COUNT(G{row}:I{row})=0,"",'
        f'"$"&FIXED(MIN(G{row}:I{row}),2,TRUE)&"-$"&FIXED(MAX(G{row}:I{row}),2,TRUE)),"")'
    )


def _scenario_bridge_incremental_formula(row: int, spec: _ScenarioDriverBridgeSpec) -> str:
    if spec.explicit_incremental:
        return f'=IFERROR(IF(C{row}="",0,C{row}),0)'
    if spec.reverse_incremental:
        return f'=IFERROR(IF(OR(B{row}="Unknown",C{row}=""),0,B{row}-C{row}),0)'
    return f'=IFERROR(IF(OR(B{row}="Unknown",C{row}=""),0,C{row}-B{row}),0)'


def _scenario_bridge_same_impact_formula(row: int) -> str:
    return f'=IFERROR(IF(OR(B{row}="Unknown",D{row}=""),0,D{row}),0)'


def _scenario_bridge_negative_impact_formula(row: int) -> str:
    return f'=IFERROR(IF(OR(B{row}="Unknown",C{row}=""),0,-(C{row}-B{row})),0)'


def _scenario_bridge_eps_manual_required(spec: _ScenarioDriverBridgeSpec) -> bool:
    if spec.eps_impact == "manual_required":
        return True
    if spec.eps_impact in {"none", "direct_per_share", "share_count", "after_tax", "earnings_conversion"}:
        return False
    return spec.driver_type in {
        SCENARIO_DRIVER_REVENUE_VOLUME,
        SCENARIO_DRIVER_MARGIN_EBITDA,
        SCENARIO_DRIVER_CAPITAL_STRUCTURE_INTEREST,
        SCENARIO_DRIVER_TAX_CREDIT_SUBSIDY,
        SCENARIO_DRIVER_MANUAL_INCREMENTAL,
    }


def _scenario_bridge_tax_conversion_label(
    tax_treatment: str,
    *,
    after_tax_factor: Optional[float] = None,
    tax_rate_ref: str = "",
    tax_source_basis: str = "",
) -> str:
    treatment = str(tax_treatment or "").strip()
    if treatment == SCENARIO_TAX_TAXABLE:
        if tax_rate_ref:
            return "active scenario tax rate; default scenario tax rate if no clean source"
        if after_tax_factor is None:
            return "No valid tax conversion"
        tax_rate = max(0.0, min(1.0, 1.0 - float(after_tax_factor)))
        return f"{tax_rate * 100:.1f}% tax rate / {float(after_tax_factor) * 100:.1f}% after-tax"
    if treatment in {SCENARIO_TAX_NON_TAXABLE, SCENARIO_TAX_NON_TAXABLE_CREDIT}:
        return "100% conversion"
    if treatment in {SCENARIO_TAX_CASH_ONLY, SCENARIO_TAX_NO_EPS_IMPACT}:
        return "n/a"
    if treatment == SCENARIO_TAX_DIRECT_EPS:
        return "direct EPS/share"
    if treatment == SCENARIO_TAX_UNKNOWN_MANUAL_REQUIRED:
        return "Manual-required"
    return tax_source_basis or "n/a"


def _scenario_bridge_default_eps_rule(
    spec: _ScenarioDriverBridgeSpec,
    *,
    after_tax_factor: Optional[float] = None,
    tax_rate_ref: str = "",
) -> str:
    if spec.eps_impact_rule:
        return spec.eps_impact_rule
    if spec.eps_impact == "share_count":
        return "EPS affected through diluted shares; no direct earnings impact"
    if spec.eps_impact == "direct_per_share":
        return "direct EPS/share delta"
    if spec.eps_impact == "none" or spec.tax_treatment == SCENARIO_TAX_NO_EPS_IMPACT:
        return "no EPS impact"
    if spec.tax_treatment == SCENARIO_TAX_CASH_ONLY:
        return "no EPS impact; cash flow only"
    if spec.tax_treatment in {SCENARIO_TAX_NON_TAXABLE, SCENARIO_TAX_NON_TAXABLE_CREDIT}:
        return "incremental / diluted shares"
    if spec.tax_treatment == SCENARIO_TAX_TAXABLE:
        return (
            "incremental * (1 - tax rate) / diluted shares"
            if after_tax_factor is not None or tax_rate_ref
            else "Manual-required until valid tax conversion exists"
        )
    return "Manual-required"


def _scenario_bridge_tax_audit_rows(
    ticker: Any,
    specs: Sequence[_ScenarioDriverBridgeSpec],
    *,
    after_tax_factor: Optional[float] = None,
    tax_rate_ref: str = "",
    tax_source_basis: str = "",
) -> List[List[Any]]:
    ticker_txt = str(ticker or "").strip().upper()
    rows: List[List[Any]] = []
    for spec in specs:
        basis = spec.tax_source_basis or tax_source_basis or "Scenario bridge classification"
        rows.append(
            [
                ticker_txt,
                spec.label,
                spec.driver_type,
                spec.tax_treatment,
                _scenario_bridge_tax_conversion_label(
                    spec.tax_treatment,
                    after_tax_factor=after_tax_factor,
                    tax_rate_ref=tax_rate_ref,
                    tax_source_basis=basis,
                ),
                basis,
                _scenario_bridge_default_eps_rule(spec, after_tax_factor=after_tax_factor, tax_rate_ref=tax_rate_ref),
                spec.audit_notes,
            ]
        )
    return rows


def write_scenario_bridge_tax_treatment_sheet(
    deps: InvestmentCaseScenarioRenderDeps,
    *,
    ticker: Any,
    specs: Sequence[_ScenarioDriverBridgeSpec],
    after_tax_factor: Optional[float] = None,
    tax_rate_ref: str = "",
    tax_source_basis: str = "",
) -> None:
    runtime = deps.runtime
    wb = runtime["wb"]
    sheet_name = "Scenario_Bridge_Tax_Treatment"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ticker_txt = str(ticker or "").strip().upper()
    preferred_after = None
    data_sheet = f"{ticker_txt}_Investment_Case_Data" if ticker_txt else ""
    case_sheet = f"{ticker_txt}_Investment_Case" if ticker_txt else ""
    if data_sheet in wb.sheetnames:
        preferred_after = wb.sheetnames.index(data_sheet) + 1
    elif case_sheet in wb.sheetnames:
        preferred_after = wb.sheetnames.index(case_sheet) + 1
    ws = wb.create_sheet(sheet_name, index=preferred_after if preferred_after is not None else None)
    headers = [
        "Ticker",
        "Bridge item",
        "Driver type",
        "Tax treatment",
        "Tax rate / conversion used",
        "Source / basis",
        "EPS impact rule",
        "Notes",
    ]
    header_fill = PatternFill("solid", fgColor="EAF3F8")
    thin = Border(
        left=Side(style="thin", color="D9E2EA"),
        right=Side(style="thin", color="D9E2EA"),
        top=Side(style="thin", color="D9E2EA"),
        bottom=Side(style="thin", color="D9E2EA"),
    )
    for cc, label in enumerate(headers, start=1):
        cell = ws.cell(1, cc, label)
        cell.fill = header_fill
        cell.font = Font(bold=True, size=11, color="1F2933")
        cell.border = thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for rr, row_vals in enumerate(
        _scenario_bridge_tax_audit_rows(
            ticker_txt,
            specs,
            after_tax_factor=after_tax_factor,
            tax_rate_ref=tax_rate_ref,
            tax_source_basis=tax_source_basis,
        ),
        start=2,
    ):
        for cc, value in enumerate(row_vals, start=1):
            cell = ws.cell(rr, cc, value)
            cell.border = thin
            cell.font = Font(size=10, color="1F2933")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[rr].height = 30
    widths = {1: 10, 2: 34, 3: 28, 4: 22, 5: 26, 6: 42, 7: 42, 8: 32}
    for cc, width in widths.items():
        ws.column_dimensions[get_column_letter(cc)].width = width
    ws.freeze_panes = "A2"


def _scenario_bridge_row_values(
    spec: _ScenarioDriverBridgeSpec,
    row: int,
    *,
    active_eps_ref: str,
    active_ebitda_ref: str = "",
    active_shares_ref: str,
    baseline_shares_ref: str,
    eps_override_ref: str,
    after_tax_factor: Optional[float] = None,
    tax_rate_ref: str = "",
) -> Tuple[Any, Any, Any, Any, Any, Any, Any, str]:
    incremental = _scenario_bridge_incremental_formula(row, spec)

    if spec.ebitda_impact == "none":
        ebitda_impact: Any = 0
    elif spec.ebitda_impact == "same" or (
        spec.ebitda_impact == "auto"
        and (
            spec.driver_type == SCENARIO_DRIVER_MARGIN_EBITDA
            or (spec.driver_type == SCENARIO_DRIVER_TAX_CREDIT_SUBSIDY and spec.subsidy_basis == "ebitda_like")
            or (spec.driver_type == SCENARIO_DRIVER_MANUAL_INCREMENTAL and spec.explicit_incremental)
        )
    ):
        ebitda_impact = _scenario_bridge_same_impact_formula(row)
    else:
        ebitda_impact = 0

    if spec.fcf_impact == "none":
        fcf_impact: Any = 0
    elif spec.fcf_impact == "negative":
        fcf_impact = _scenario_bridge_negative_impact_formula(row)
    elif spec.fcf_impact == "same" or (
        spec.fcf_impact == "auto"
        and spec.driver_type in {SCENARIO_DRIVER_CAPITAL_STRUCTURE_INTEREST}
    ):
        fcf_impact = _scenario_bridge_same_impact_formula(row)
    else:
        fcf_impact = 0

    if spec.eps_impact == "none":
        eps_impact: Any = 0
    elif spec.eps_impact == "direct_per_share":
        eps_impact = f'=IFERROR(IF(D{row}="","",D{row}),"")'
    elif spec.eps_impact == "share_count":
        eps_impact = (
            f'=IFERROR(IF(OR({active_eps_ref}="",'
            f'B{row}="Unknown",C{row}="",C{row}=0),"",'
            f'{active_eps_ref}*(B{row}/C{row})-{active_eps_ref}),"")'
        )
    elif spec.tax_treatment in {SCENARIO_TAX_NON_TAXABLE, SCENARIO_TAX_NON_TAXABLE_CREDIT}:
        eps_impact = (
            f'=IFERROR(IF(OR({active_shares_ref}="",{active_shares_ref}=0,D{row}=""),"",'
            f'D{row}/{active_shares_ref}),"")'
        )
    elif spec.tax_treatment == SCENARIO_TAX_TAXABLE and tax_rate_ref:
        eps_impact = (
            f'=IFERROR(IF(OR({tax_rate_ref}="",{tax_rate_ref}<0,{tax_rate_ref}>0.35,'
            f'{active_shares_ref}="",{active_shares_ref}=0,D{row}=""),'
            f'"Manual-required",D{row}*(1-{tax_rate_ref})/{active_shares_ref}),"Manual-required")'
        )
    elif spec.tax_treatment == SCENARIO_TAX_TAXABLE and after_tax_factor is not None:
        eps_impact = (
            f'=IFERROR(IF(OR({active_shares_ref}="",'
            f'{active_shares_ref}=0,D{row}=""),"",D{row}*{float(after_tax_factor):.6f}/{active_shares_ref}),"")'
        )
    elif spec.tax_treatment == SCENARIO_TAX_TAXABLE:
        eps_impact = "Manual-required"
    elif spec.tax_treatment in {SCENARIO_TAX_CASH_ONLY, SCENARIO_TAX_NO_EPS_IMPACT}:
        eps_impact = 0
    elif _scenario_bridge_eps_manual_required(spec):
        eps_impact = "Manual-required"
    else:
        eps_impact = 0

    return (
        spec.label,
        spec.baseline,
        spec.active,
        incremental,
        eps_impact,
        ebitda_impact,
        fcf_impact,
        spec.read,
    )


def _scenario_bridge_eps_value_formula(
    *,
    summary_row: int,
    eps_override_ref: str,
    active_eps_ref: str,
    active_shares_ref: str,
    baseline_shares_ref: str,
    eps_impact_start: int,
    eps_impact_end: int,
) -> str:
    return (
        f'=IFERROR(IF(B{summary_row}="","",'
        f'IF({eps_override_ref}<>"",B{summary_row},'
        f'IF(AND({active_eps_ref}<>"",{baseline_shares_ref}<>"",{active_shares_ref}<>"",{active_shares_ref}<>0),'
        f'({active_eps_ref}*{baseline_shares_ref}+SUM(E{eps_impact_start}:E{eps_impact_end})*{active_shares_ref})/{active_shares_ref},'
        f'B{summary_row}+SUM(E{eps_impact_start}:E{eps_impact_end})))),"")'
    )
