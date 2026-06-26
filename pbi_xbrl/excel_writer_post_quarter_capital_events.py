"""Workbook surfaces for normalized post-quarter capital events."""
from __future__ import annotations

from typing import Any, Callable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def write_post_quarter_capital_events_sheet(
    wb: Any,
    events: pd.DataFrame,
) -> None:
    if not isinstance(events, pd.DataFrame) or events.empty:
        return
    if "PostQuarter_Capital_Events" in wb.sheetnames:
        del wb["PostQuarter_Capital_Events"]
    ws = wb.create_sheet("PostQuarter_Capital_Events")
    headers = [str(column) for column in events.columns]
    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, record in enumerate(events.to_dict("records"), start=2):
        for column_index, header in enumerate(headers, start=1):
            value = record.get(header)
            if pd.isna(value):
                value = None
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    end_col = ws.cell(row=1, column=len(headers)).column_letter
    ws.freeze_panes = "A2"
    for column_index, header in enumerate(headers, start=1):
        width = 18
        if header in {"source_documents", "source_paths", "source_urls", "used_surfaces"}:
            width = 48
        ws.column_dimensions[ws.cell(row=1, column=column_index).column_letter].width = width


def render_gpre_warrant_valuation_overlay(
    *,
    ws: Any,
    start_row: int,
    events: pd.DataFrame,
    section_fill: Any,
    bold: Any,
    row_fill: Callable[[int, Any], None],
    set_cell_comment: Callable[..., None],
) -> int:
    if not isinstance(events, pd.DataFrame) or events.empty:
        return start_row
    event_rows = events[
        events.get("event_type", pd.Series(dtype=object)).astype(str).eq("warrant_dilution")
    ]
    if event_rows.empty:
        return start_row
    event = event_rows.iloc[-1]
    panel_label_start = 19
    panel_label_end = 22
    panel_value_start = 23
    panel_value_end = 26
    row = start_row
    try:
        ws.merge_cells(
            start_row=row,
            start_column=panel_label_start,
            end_row=row,
            end_column=panel_value_end,
        )
    except Exception:
        pass
    ws.cell(
        row=row,
        column=panel_label_start,
        value="Post-quarter warrant dilution overlay",
    ).font = bold
    for column in range(panel_label_start, panel_value_end + 1):
        ws.cell(row=row, column=column).fill = section_fill
    row += 1
    narrative = (
        "Post-quarter BlackRock warrant overlay: 500k warrants issued; "
        "S-3 registers up to 550k common shares issuable on exercise. "
        "Reported 2026-Q1 shares/EPS unchanged; valuation full-dilution "
        "sensitivity uses +0.55m shares."
    )
    ws.cell(row=row, column=panel_label_start, value=narrative)
    try:
        ws.merge_cells(
            start_row=row,
            start_column=panel_label_start,
            end_row=row,
            end_column=panel_value_end,
        )
    except Exception:
        pass
    ws.cell(row=row, column=panel_label_start).alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )
    source_note = (
        f"Source: {event.get('filing_type')} accession {event.get('accession')}\n"
        f"{event.get('source_paths')}"
    )
    try:
        set_cell_comment(ws.cell(row=row, column=panel_label_start), source_note)
    except Exception:
        pass
    row += 1
    values = (
        ("Warrants issued (m)", float(event["warrants_issued"]) / 1e6, "#,##0.000"),
        (
            "Maximum common shares issuable (m)",
            float(event["potential_common_shares_issuable_max"]) / 1e6,
            "#,##0.000",
        ),
        ("Exercise price", float(event["exercise_price"]), "$#,##0.00"),
        ("Expiration", str(event["expiration_date"]), "yyyy-mm-dd"),
        (
            "Beneficial ownership limitation",
            float(event["beneficial_ownership_limitation"]),
            "0.0%",
        ),
        ("Reported diluted shares (m)", "=SharesDiluted", "#,##0.000"),
        (
            "Post-quarter potential dilution shares (m)",
            float(event["potential_common_shares_issuable_max"]) / 1e6,
            "#,##0.000",
        ),
        ("Full-dilution overlay shares (m)", "=SharesDiluted+0.550", "#,##0.000"),
        (
            "Value/share sensitivity: diluted + post-quarter warrants",
            '=IF(OR(SharesDiluted="",Adj_EBITDA=""),"",'
            "(Target_EV_AdjEBITDA*Adj_EBITDA-NetDebt)/(SharesDiluted+0.550))",
            "$#,##0.00",
        ),
    )
    for label, value, number_format in values:
        try:
            ws.merge_cells(
                start_row=row,
                start_column=panel_label_start,
                end_row=row,
                end_column=panel_label_end,
            )
            ws.merge_cells(
                start_row=row,
                start_column=panel_value_start,
                end_row=row,
                end_column=panel_value_end,
            )
        except Exception:
            pass
        ws.cell(row=row, column=panel_label_start, value=label)
        ws.cell(row=row, column=panel_value_start, value=value).number_format = number_format
        ws.cell(row=row, column=panel_label_start).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        ws.cell(row=row, column=panel_value_start).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        row += 1
    for column in range(panel_label_start, panel_label_end + 1):
        letter = get_column_letter(column)
        ws.column_dimensions[letter].width = max(
            float(ws.column_dimensions[letter].width or 0),
            14,
        )
    for column in range(panel_value_start, panel_value_end + 1):
        letter = get_column_letter(column)
        ws.column_dimensions[letter].width = max(
            float(ws.column_dimensions[letter].width or 0),
            12,
        )
    return row
