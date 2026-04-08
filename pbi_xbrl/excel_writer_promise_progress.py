"""Dedicated writer surface for the Promise_Progress_UI sheet."""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from datetime import date
from typing import Any, Callable, Dict, List, Sequence

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class PromiseProgressSheetInputs:
    wb: Any
    sheet_name: str
    quarters: List[date]
    rows_by_quarter: Dict[date, List[Dict[str, Any]]]
    generated_at_text: str
    pp_rationale_col_width_default: float
    empty_message: str = ""


@dataclass(frozen=True)
class PromiseProgressRenderHelpers:
    write_analysis_sheet_title_and_metadata: Callable[..., None]
    render_stacked_quarter_blocks: Callable[..., int]
    row_writer: Callable[[Any, int, date, Dict[str, Any]], None]
    get_analysis_sheet_style_bundle: Callable[[], Dict[str, Any]]
    estimate_wrapped_line_count: Callable[..., float]
    parse_dollar_amount: Callable[[Any], Any]
    post_render_cleanups: Sequence[Callable[[Any], None]] = field(default_factory=tuple)


@dataclass(frozen=True)
class PromiseProgressSheetResult:
    ws: Any


def _block_header_writer(
    ws: Any,
    row_idx: int,
    max_col: int,
    *,
    helpers: PromiseProgressRenderHelpers,
) -> int:
    theme = helpers.get_analysis_sheet_style_bundle()
    hdr_fill = copy(theme["header_fill"])
    thin_border = copy(theme["thin_border"])
    labels = {
        1: "Metric",
        2: "Target",
        3: "Latest",
        4: "Result",
        5: "Rationale",
        6: "Stated",
        7: "Last Seen",
        8: "Carried To",
        9: "Evaluated Through",
        10: "Evidence",
        15: "Promise Id",
    }
    for cc in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=cc, value=labels.get(cc, ""))
        cell.font = Font(bold=True, size=11, color=str(theme["text_dark"]))
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[row_idx].height = 16.0
    return row_idx + 1


def _promise_sheet_metric_is_monetary(metric_text: Any) -> bool:
    metric_low = str(metric_text or "").strip().lower()
    return any(
        token in metric_low
        for token in (
            "revenue",
            "ebit",
            "ebitda",
            "eps",
            "fcf",
            "cost savings",
            "liquidity",
            "interest expense",
            "45z",
            "debt",
            "buyback",
            "dividend",
        )
    )


def _promise_sheet_moneyish(value_in: Any, *, helpers: PromiseProgressRenderHelpers) -> bool:
    txt = str(value_in or "").strip()
    if not txt:
        return False
    if helpers.parse_dollar_amount(txt) is not None:
        return True
    return txt[:1] in {">", "<", "~"} and "$" in txt or txt.startswith("$")


def _apply_promise_progress_visible_formatting(
    ws: Any,
    *,
    helpers: PromiseProgressRenderHelpers,
    pp_rationale_col_width_default: float,
) -> None:
    theme = helpers.get_analysis_sheet_style_bundle()
    zebra_fills = [copy(theme["neutral_fill_alt"]), copy(theme["neutral_fill"])]

    def _apply_result_fill_local(cell_in: Any) -> None:
        result_low = str(getattr(cell_in, "value", "") or "").strip().lower()
        fill_color = "E7EDF3"
        if result_low == "missed":
            fill_color = "F4CCCC"
        elif result_low == "at risk":
            fill_color = "FCE5CD"
        elif result_low == "beat":
            fill_color = "A9D18E"
        elif result_low == "hit":
            fill_color = "C6EFCE"
        elif result_low == "completed":
            fill_color = "70AD47"
        elif result_low == "on track":
            fill_color = "D9EAD3"
        elif result_low == "updated":
            fill_color = "D9EAF7"
        elif result_low == "open":
            fill_color = "E7EDF3"
        elif result_low == "not yet measurable":
            fill_color = "E7E6E6"
        cell_in.fill = PatternFill("solid", fgColor=fill_color)

    blank_side_local = Side(style=None)
    zebra_idx = 0
    for rr in range(1, ws.max_row + 1):
        metric_val = str(ws.cell(rr, 1).value or "").strip()
        row_vals = [str(ws.cell(rr, cc).value or "").strip() for cc in range(1, 16)]
        if metric_val == "Promise Progress":
            for cc in range(1, 16):
                cell = ws.cell(rr, cc)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
                cell.border = copy(theme["thin_border"])
            continue
        if metric_val.startswith("Promise progress (As of "):
            continue
        if metric_val.startswith("Generated at "):
            for cc in range(1, 16):
                cell = ws.cell(rr, cc)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=blank_side_local,
                    bottom=blank_side_local,
                )
            ws.cell(rr, 4).fill = PatternFill(fill_type=None)
            continue
        if not metric_val:
            continue
        if metric_val == "Metric":
            for cc in range(1, 16):
                ws.cell(rr, cc).alignment = Alignment(
                    horizontal="right" if cc in {2, 3} else "left",
                    vertical="center",
                    wrap_text=False,
                )
            continue
        if metric_val and not any(row_vals[1:]):
            for cc in range(1, 16):
                cell = ws.cell(rr, cc)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            continue
        metric_cell = ws.cell(rr, 1)
        target_cell = ws.cell(rr, 2)
        latest_cell = ws.cell(rr, 3)
        result_cell = ws.cell(rr, 4)
        rationale_cell = ws.cell(rr, 5)
        meta_cols = [6, 7, 8, 9, 10, 15]

        metric_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        target_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        latest_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        result_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        _apply_result_fill_local(result_cell)
        rationale_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for cc in meta_cols:
            ws.cell(rr, cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        for cc in range(11, 15):
            ws.cell(rr, cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        if _promise_sheet_metric_is_monetary(metric_val):
            try:
                latest_num = float(latest_cell.value)
            except Exception:
                latest_num = None
            if latest_num is not None:
                metric_low = metric_val.lower()
                if "eps" in metric_low:
                    latest_cell.number_format = "$0.00"
                elif abs(latest_num) >= 1_000_000:
                    latest_cell.number_format = '$#,##0.000,,"m"'
                elif abs(latest_num) >= 1_000:
                    latest_cell.number_format = "$#,##0.0"
                else:
                    latest_cell.number_format = "$0.00"
        rationale_txt = str(rationale_cell.value or "").strip()
        rationale_lines = helpers.estimate_wrapped_line_count(
            rationale_txt,
            float(ws.column_dimensions["E"].width or pp_rationale_col_width_default),
            min_lines=1,
            max_lines=4,
        )
        preferred_height = 20.0 if rationale_lines <= 1.25 else (26.0 if rationale_lines <= 2.1 else 32.0)
        current_height = float(ws.row_dimensions[rr].height or preferred_height)
        ws.row_dimensions[rr].height = max(preferred_height, min(current_height, 40.0))
        for cc in range(1, 16):
            cell = ws.cell(rr, cc)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=blank_side_local,
                bottom=blank_side_local,
            )
            if cc != 4:
                cell.fill = copy(zebra_fills[zebra_idx % 2])
        zebra_idx += 1


def _px_to_excel_width(px: float) -> float:
    try:
        value = (float(px) - 5.0) / 7.0
    except Exception:
        value = 10.0
    return max(0.1, value)


def write_promise_progress_sheet(
    inputs: PromiseProgressSheetInputs,
    helpers: PromiseProgressRenderHelpers,
) -> PromiseProgressSheetResult:
    ws = inputs.wb.create_sheet(inputs.sheet_name)
    helpers.write_analysis_sheet_title_and_metadata(
        ws,
        "Promise Progress",
        inputs.generated_at_text,
        max_col=15,
    )
    if inputs.empty_message:
        ws["A3"] = inputs.empty_message
        ws.freeze_panes = "A3"
        return PromiseProgressSheetResult(ws=ws)

    helpers.render_stacked_quarter_blocks(
        ws=ws,
        quarters=inputs.quarters,
        rows_by_quarter=inputs.rows_by_quarter,
        max_col=15,
        block_title_fn=lambda qd: f"Promise progress (As of {qd})",
        row_writer=helpers.row_writer,
        block_header_writer=lambda sheet, row_idx, _qd, max_col: _block_header_writer(
            sheet,
            row_idx,
            max_col,
            helpers=helpers,
        ),
        start_row=3,
        blank_row_between=False,
    )

    for cleanup in helpers.post_render_cleanups:
        cleanup(ws)

    _apply_promise_progress_visible_formatting(
        ws,
        helpers=helpers,
        pp_rationale_col_width_default=inputs.pp_rationale_col_width_default,
    )
    ws.row_dimensions[1].height = 27.0

    for rr in range(2, ws.max_row + 1):
        row_vals = [str(ws.cell(row=rr, column=cc).value or "").strip() for cc in range(1, 16)]
        if not any(row_vals):
            ws.row_dimensions[rr].height = 12.0
            continue
        if row_vals[0].startswith("Promise progress (As of "):
            ws.row_dimensions[rr].height = 19.5
            continue
        if row_vals[0] and not any(row_vals[1:]):
            ws.row_dimensions[rr].height = 16.0
            continue
        if row_vals[0].lower() == "metric" and row_vals[14].lower() == "promise id":
            ws.row_dimensions[rr].height = 19.5
            continue
        if row_vals[0].lower().startswith("guidance accuracy"):
            ws.row_dimensions[rr].height = 16.0

    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = _px_to_excel_width(272.0)
    ws.column_dimensions["C"].width = _px_to_excel_width(272.0)
    ws.column_dimensions["D"].width = 17
    ws.column_dimensions["E"].width = _px_to_excel_width(554.0)
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 4
    ws.column_dimensions["L"].width = 4
    ws.column_dimensions["M"].width = 4
    ws.column_dimensions["N"].width = 4
    ws.column_dimensions["O"].width = 24
    return PromiseProgressSheetResult(ws=ws)
