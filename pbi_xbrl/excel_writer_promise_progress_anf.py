"""ANF-specific Promise Progress writer."""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from typing import Any, Callable, Dict, List, Mapping, Sequence

from openpyxl.styles import Alignment, Font, PatternFill


@dataclass(frozen=True)
class AnfPromiseProgressWriterDeps:
    wb: Any
    slides_guidance: Any
    hist: Any
    generated_at_text: str
    promise_visible_max_col: int
    promise_timeline_headers: Sequence[str]
    write_analysis_sheet_title_and_metadata: Callable[..., None]
    get_analysis_sheet_style_bundle: Callable[[], Mapping[str, Any]]
    anf_build_promise_progress_sections: Callable[..., Mapping[str, Any]]
    management_credibility_scorecard_rows: Callable[..., Sequence[Any]]
    anf_clean_visible_ui_text: Callable[..., str]


def write_anf_promise_progress_ui_sheet(
    deps: AnfPromiseProgressWriterDeps,
) -> List[Dict[str, Any]]:
    qa_rows: List[Dict[str, Any]] = []
    ws = deps.wb.create_sheet("Promise_Progress_UI")
    ws.sheet_view.zoomScale = 112
    promise_max_col = deps.promise_visible_max_col
    deps.write_analysis_sheet_title_and_metadata(
        ws,
        "Promise Progress",
        deps.generated_at_text.replace("Quarter blocks", "ANF guidance tracker"),
        max_col=promise_max_col,
    )
    theme = deps.get_analysis_sheet_style_bundle()
    section_fill = PatternFill("solid", fgColor="5B9BD5")
    header_fill = copy(theme["header_fill"])
    thin_border = copy(theme["thin_border"])
    neutral_fill = copy(theme["neutral_fill"])
    neutral_alt = copy(theme["neutral_fill_alt"])
    text_dark = str(theme["text_dark"])
    text_muted = str(theme["text_muted"])

    def _section_bar(row_idx: int, title: str) -> int:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=promise_max_col)
        cell = ws.cell(row=row_idx, column=1, value=title)
        cell.font = Font(bold=True, size=13, color="FFFFFF")
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for cc in range(1, promise_max_col + 1):
            ws.cell(row=row_idx, column=cc).fill = section_fill
            ws.cell(row=row_idx, column=cc).border = thin_border
        ws.row_dimensions[row_idx].height = 24.0
        return row_idx + 1

    def _write_header(row_idx: int, labels: Sequence[str]) -> int:
        for cc in range(1, promise_max_col + 1):
            value = labels[cc - 1] if cc <= len(labels) else ""
            cell = ws.cell(row=row_idx, column=cc, value=value)
            cell.font = Font(bold=True, size=11, color=text_dark)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        ws.row_dimensions[row_idx].height = 20.0
        return row_idx + 1

    def _status_fill(status: str) -> PatternFill:
        low = str(status or "").strip().lower()
        if low in {"completed", "achieved"}:
            return PatternFill("solid", fgColor="009E73")
        if low in {"met", "hit"}:
            return PatternFill("solid", fgColor="66C2A5")
        if low in {"on track", "on_track"}:
            return PatternFill("solid", fgColor="56B4E9")
        if low == "open":
            return PatternFill("solid", fgColor="A6CEE3")
        if low == "mixed":
            return PatternFill("solid", fgColor="E69F00")
        if low == "met-ish":
            return PatternFill("solid", fgColor="F0E442")
        if low == "basis-dependent":
            return PatternFill("solid", fgColor="CC79A7")
        if low in {"fail", "miss", "missed"}:
            return PatternFill("solid", fgColor="D55E00")
        if low in {"n/a", "na", "not applicable"}:
            return PatternFill("solid", fgColor="D9D9D9")
        return copy(neutral_fill)

    sections = deps.anf_build_promise_progress_sections(deps.slides_guidance, deps.hist)
    row_idx = 3
    row_idx = _section_bar(row_idx, "Management Credibility Scorecard")
    header_row = row_idx
    row_idx = _write_header(row_idx, ["Category", "Score", "Evidence", "", "", "", "Read"])
    ws.merge_cells(start_row=header_row, start_column=3, end_row=header_row, end_column=6)
    ws.merge_cells(start_row=header_row, start_column=7, end_row=header_row, end_column=promise_max_col)
    for idx, (category, score, evidence, read) in enumerate(deps.management_credibility_scorecard_rows("ANF")):
        data_row = row_idx
        vals = [category, score, evidence, "", "", "", read]
        fill = copy(neutral_alt if idx % 2 == 0 else neutral_fill)
        for cc in range(1, promise_max_col + 1):
            value = vals[cc - 1] if cc <= len(vals) else ""
            cell = ws.cell(row=row_idx, column=cc, value=deps.anf_clean_visible_ui_text(value, max_chars=260))
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(size=11, color=text_dark)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc in {3, 7})
        ws.merge_cells(start_row=data_row, start_column=3, end_row=data_row, end_column=6)
        ws.merge_cells(start_row=data_row, start_column=7, end_row=data_row, end_column=promise_max_col)
        ws.row_dimensions[row_idx].height = 24.0
        row_idx += 1
    row_idx += 1
    row_idx = _section_bar(row_idx, "2025 guidance progression")
    header_row = row_idx
    row_idx = _write_header(
        row_idx,
        ["Metric", "Initial guide", "Q1 update", "Q2 update", "Q3 update", "Jan 2026 update", "Actual", "Status", "Notes/source"],
    )
    ws.merge_cells(start_row=header_row, start_column=9, end_row=header_row, end_column=10)
    zebra = 0
    for rec in sections.get("2025 guidance progression", []):
        data_row = row_idx
        vals = [
            rec.get("Metric", ""),
            rec.get("Initial guide", ""),
            rec.get("Q1 update", ""),
            rec.get("Q2 update", ""),
            rec.get("Q3 update", ""),
            rec.get("Jan 2026 update", ""),
            rec.get("Actual", ""),
            rec.get("Status", ""),
            rec.get("Notes/source", ""),
        ]
        fill = copy(neutral_alt if zebra % 2 == 0 else neutral_fill)
        for cc, value in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=cc, value=deps.anf_clean_visible_ui_text(value, max_chars=260))
            cell.fill = _status_fill(value) if cc == 8 else fill
            cell.border = thin_border
            cell.font = Font(size=11, color=text_dark)
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=cc == 9,
            )
        ws.merge_cells(start_row=data_row, start_column=9, end_row=data_row, end_column=10)
        ws.cell(row=data_row, column=9).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_idx].height = 24.0
        row_idx += 1
        zebra += 1

    for older_section in (
        "2024 guidance progression",
        "2023 guidance progression",
        "2022 guidance progression",
    ):
        rows_for_section = list(sections.get(older_section, []) or [])
        if not rows_for_section:
            continue
        row_idx += 1
        row_idx = _section_bar(row_idx, older_section)
        header_row = row_idx
        row_idx = _write_header(
            row_idx,
            ["Metric", "Initial guide", "Q1 update", "Q2 update", "Q3 update", "Q4 update", "Actual", "Status", "Notes/source"],
        )
        ws.merge_cells(start_row=header_row, start_column=9, end_row=header_row, end_column=10)
        for rec in rows_for_section:
            data_row = row_idx
            vals = [
                rec.get("Metric", ""),
                rec.get("Initial guide", ""),
                rec.get("Q1 update", ""),
                rec.get("Q2 update", ""),
                rec.get("Q3 update", ""),
                rec.get("Q4 update", ""),
                rec.get("Actual", ""),
                rec.get("Status", ""),
                rec.get("Notes/source", ""),
            ]
            fill = copy(neutral_alt if zebra % 2 == 0 else neutral_fill)
            for cc, value in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=cc, value=deps.anf_clean_visible_ui_text(value, max_chars=260))
                cell.fill = _status_fill(value) if cc == 8 else fill
                cell.border = thin_border
                cell.font = Font(size=11, color=text_dark)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc == 9)
            ws.merge_cells(start_row=data_row, start_column=9, end_row=data_row, end_column=10)
            ws.cell(row=data_row, column=9).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[row_idx].height = 24.0
            row_idx += 1
            zebra += 1

    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=promise_max_col)
    cred_cell = ws.cell(
        row=row_idx,
        column=1,
        value="Guidance credibility read: management delivered sales and buybacks, but margin/EPS need GAAP-vs-adjusted basis discipline.",
    )
    cred_cell.fill = PatternFill("solid", fgColor="D9EAF7")
    cred_cell.font = Font(bold=True, size=11, color=text_dark)
    cred_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for cc in range(1, promise_max_col + 1):
        ws.cell(row=row_idx, column=cc).fill = PatternFill("solid", fgColor="D9EAF7")
        ws.cell(row=row_idx, column=cc).border = thin_border
    ws.row_dimensions[row_idx].height = 22.0
    row_idx += 1
    row_idx += 1
    row_idx = _section_bar(row_idx, "2026 open guidance")
    header_row = row_idx
    row_idx = _write_header(row_idx, ["Metric", "Current guide", "Horizon", "Status", "Notes/source"])
    ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=10)
    previous_open_horizon = ""
    for rec in sections.get("2026 open guidance", []):
        current_open_horizon = str(rec.get("Horizon") or "").strip()
        if previous_open_horizon and current_open_horizon and current_open_horizon != previous_open_horizon:
            ws.row_dimensions[row_idx].height = 8.0
            row_idx += 1
        data_row = row_idx
        vals = [
            rec.get("Metric", ""),
            rec.get("Current guide", ""),
            rec.get("Horizon", ""),
            rec.get("Status", ""),
            rec.get("Notes/source", ""),
        ]
        fill = copy(neutral_alt if zebra % 2 == 0 else neutral_fill)
        for cc in range(1, promise_max_col + 1):
            value = vals[cc - 1] if cc <= len(vals) else ""
            cell = ws.cell(row=row_idx, column=cc, value=deps.anf_clean_visible_ui_text(value, max_chars=260))
            cell.fill = _status_fill(value) if cc == 4 else fill
            cell.border = thin_border
            cell.font = Font(size=11, color=text_dark)
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=cc == 5,
            )
        ws.merge_cells(start_row=data_row, start_column=5, end_row=data_row, end_column=10)
        ws.cell(row=data_row, column=5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_idx].height = 23.0
        row_idx += 1
        zebra += 1
        previous_open_horizon = current_open_horizon

    row_idx += 1
    row_idx = _section_bar(row_idx, "Quarterly guidance timeline / revision log")
    timeline_headers = [
        *deps.promise_timeline_headers,
        "",
    ]
    last_timeline_group = None
    for rec in sections.get("Quarterly guidance timeline / revision log", []):
        stated_group = deps.anf_clean_visible_ui_text(rec.get("Stated in", "") or "Timeline")
        if stated_group != last_timeline_group:
            if last_timeline_group is not None:
                row_idx += 1
            row_idx = _section_bar(row_idx, f"{stated_group} revisions")
            row_idx = _write_header(row_idx, timeline_headers)
            last_timeline_group = stated_group
        source_date_txt = str(rec.get("Source date / source quarter") or "")
        source_date_txt = source_date_txt.split("/", 1)[0].strip() if source_date_txt else ""
        vals = [
            rec.get("Metric", ""),
            rec.get("Previous guide", ""),
            rec.get("New/current guide", ""),
            rec.get("Change type", ""),
            rec.get("Actual", ""),
            rec.get("Progress / run-rate", ""),
            rec.get("Status", ""),
            rec.get("Horizon", ""),
            rec.get("Stated in", ""),
            source_date_txt,
            rec.get("Source / note", ""),
        ]
        fill = copy(neutral_alt if zebra % 2 == 0 else neutral_fill)
        for cc, value in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=cc, value=deps.anf_clean_visible_ui_text(value, max_chars=220))
            cell.fill = _status_fill(value) if cc == 7 else fill
            cell.border = thin_border
            cell.font = Font(size=11, color=text_dark)
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=cc in {1, 11},
            )
        ws.row_dimensions[row_idx].height = 24.0
        row_idx += 1
        zebra += 1

    ws.cell(row=row_idx + 1, column=1, value="Labels are fiscal periods; Q4 2025 ended 2026-01-31.")
    ws.cell(row=row_idx + 1, column=1).font = Font(italic=True, size=9, color=text_muted)
    ws.freeze_panes = "A2"
    for col, width in {
        "A": 28,
        "B": 28,
        "C": 32,
        "D": 15,
        "E": 22,
        "F": 28,
        "G": 15,
        "H": 14,
        "I": 16,
        "J": 14,
        "K": 42,
        "L": 42,
    }.items():
        ws.column_dimensions[col].width = width
    return qa_rows
