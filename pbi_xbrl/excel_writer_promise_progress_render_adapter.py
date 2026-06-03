"""Promise Progress per-row render adapter."""
from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from typing import Any, Callable, Dict, Mapping, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .guidance_lexicon import normalize_text as glx_normalize_text
from .quarter_notes_lexicon import compact_snippet as qn_compact_snippet


@dataclass(frozen=True)
class PromiseProgressRowWriterDeps:
    is_pbi_profile: bool
    pp_rationale_col_width_default: float
    ev_map_q: Mapping[Tuple[str, date], int]
    ev_map_pid: Mapping[str, int]
    display_progress_metric: Callable[[Dict[str, Any]], str]
    excel_safe_text: Callable[..., Any]
    safe_cell: Callable[[Any], Any]
    set_num: Callable[[Any, Any], bool]
    short_pid: Callable[[str], str]
    parse_dollar_amount: Callable[[Any], Any]
    q_label: Callable[[Any], str]
    looks_pbi_fragment_text: Callable[[Any], bool]
    lookup_pbi_structured_guidance_target: Callable[..., Dict[str, Any]]
    extract_pbi_target_display: Callable[..., str]
    get_analysis_sheet_style_bundle: Callable[[], Dict[str, Any]]
    apply_hyperlink_look: Callable[[Any, str], None]
    set_cell_comment: Callable[..., None]
    estimate_wrapped_row_height: Callable[..., float]
    estimate_wrapped_line_count: Callable[..., float]


def build_promise_progress_row_writer(
    deps: PromiseProgressRowWriterDeps,
) -> Callable[[Any, int, date, Dict[str, Any]], None]:
    is_pbi_profile = deps.is_pbi_profile
    pp_rationale_col_width_default = deps.pp_rationale_col_width_default
    ev_map_q = deps.ev_map_q
    ev_map_pid = deps.ev_map_pid
    _display_progress_metric = deps.display_progress_metric
    _excel_safe_text_local = deps.excel_safe_text
    _safe_cell = deps.safe_cell
    _set_num = deps.set_num
    _short_pid = deps.short_pid
    _parse_dollar_amount = deps.parse_dollar_amount
    _q_label = deps.q_label
    _looks_pbi_fragment_text = deps.looks_pbi_fragment_text
    _lookup_pbi_structured_guidance_target = deps.lookup_pbi_structured_guidance_target
    _extract_pbi_target_display = deps.extract_pbi_target_display
    _get_analysis_sheet_style_bundle = deps.get_analysis_sheet_style_bundle
    _apply_hyperlink_look = deps.apply_hyperlink_look
    _set_cell_comment_local = deps.set_cell_comment
    _estimate_wrapped_row_height = deps.estimate_wrapped_row_height
    _estimate_wrapped_line_count = deps.estimate_wrapped_line_count

    def _row_writer(_ws: Any, row_idx: int, qd: date, item: Dict[str, Any]) -> None:
        row_type = str(item.get("row_type") or "").strip().lower()
        if row_type == "section":
            label = _excel_safe_text_local(str(item.get("section_label") or "Guidance accuracy"))
            theme = _get_analysis_sheet_style_bundle()
            sec_fill = copy(theme["title_fill"])
            thin_border = copy(theme["thin_border"])
            for cc in range(1, 16):
                c_sec = _ws.cell(row=row_idx, column=cc, value=label if cc == 1 else "")
                c_sec.fill = sec_fill
                c_sec.font = Font(bold=True, size=11, color="FFFFFF")
                c_sec.alignment = Alignment(horizontal="left", vertical="center")
                c_sec.border = thin_border
            _ws.row_dimensions[row_idx].height = 16.0
            return

        def _progress_metric_is_monetary(metric_text: Any) -> bool:
            metric_low = str(metric_text or "").strip().lower()
            return any(
                token in metric_low
                for token in (
                    "revenue",
                    "ebit",
                    "ebitda",
                    "eps",
                    "fcf",
                    "cost savings",
                    "liquidity",
                    "interest expense",
                    "45z",
                    "debt",
                    "buyback",
                    "dividend",
                )
            )

        def _progress_moneyish_text(text_in: Any) -> bool:
            txt = str(text_in or "").strip()
            if not txt:
                return False
            if _parse_dollar_amount(txt) is not None:
                return True
            return bool(re.match(r"^[><~]?\s*\$[0-9]", txt))

        def _apply_progress_monetary_format(cell: Any, metric_text: Any) -> None:
            if not _progress_metric_is_monetary(metric_text):
                return
            try:
                val = float(cell.value)
            except Exception:
                return
            metric_low = str(metric_text or "").strip().lower()
            if "eps" in metric_low:
                cell.number_format = "$0.00"
            elif abs(val) >= 1_000_000:
                cell.number_format = '$#,##0.000,,"m"'
            elif abs(val) >= 1_000:
                cell.number_format = "$#,##0.0"
            else:
                cell.number_format = "$0.00"

        display_metric = _display_progress_metric(item)
        pid = str(item.get("promise_id") or "")
        if is_pbi_profile and display_metric in {"Revenue guidance", "Adjusted EBIT guidance", "EPS guidance", "FCF target"}:
            pid = f"guidance:{display_metric.lower().replace(' ', '_')}"
        pid_display = pid if pid.startswith("guidance:") else _short_pid(pid)
        _ws.cell(row=row_idx, column=1, value=_excel_safe_text_local(display_metric))
        _ws.cell(row=row_idx, column=15, value=_excel_safe_text_local(pid_display))
        c_t = _ws.cell(row=row_idx, column=2)
        target_value = item.get("target")
        if (
            str(target_value or "").strip().lower() in {"1", "1.0"}
            and (
                str(item.get("promise_type") or "").strip().lower() == "milestone"
                or display_metric in {"Strategic milestone", "Advantage Nebraska startup", "45Z plant qualification readiness"}
            )
        ):
            target_value = ""
        if is_pbi_profile and (
            not str(target_value or "").strip()
            or _looks_pbi_fragment_text(target_value)
        ):
            guidance_like_metric = _display_progress_metric(item)
            structured_target = ""
            if guidance_like_metric in {"Revenue guidance", "Adjusted EBIT guidance", "EPS guidance", "FCF target"}:
                structured_guidance = _lookup_pbi_structured_guidance_target(
                    qd if isinstance(qd, date) else None,
                    guidance_like_metric,
                    " | ".join(
                        [
                            str(item.get("metric_display") or ""),
                            str(item.get("metric_ref") or ""),
                            str(item.get("rationale") or ""),
                            str(item.get("latest") or ""),
                        ]
                    ),
                )
                structured_target = str((structured_guidance or {}).get("target_display") or "").strip()
            target_value = structured_target or _extract_pbi_target_display(
                " | ".join(
                    [
                        str(item.get("metric_display") or ""),
                        str(item.get("metric_ref") or ""),
                        str(item.get("rationale") or ""),
                        str(item.get("latest") or ""),
                    ]
                ),
                guidance_like_metric,
            )
        target_is_numeric = _set_num(c_t, target_value)
        if not target_is_numeric:
            c_t.value = _safe_cell(target_value)
        else:
            _apply_progress_monetary_format(c_t, display_metric)
        target_txt = str(c_t.value or "")
        target_moneyish = _progress_moneyish_text(target_value if not target_is_numeric else c_t.value)
        c_t.alignment = Alignment(
            horizontal="right" if (target_is_numeric or target_moneyish) else "left",
            vertical="center" if (target_is_numeric or target_moneyish or len(target_txt) <= 18) else "top",
            wrap_text=(not target_is_numeric and not target_moneyish and len(target_txt) > 18),
        )
        c_l = _ws.cell(row=row_idx, column=3)
        latest_value = "" if str(item.get("latest") or "").strip().lower() == "nan" else item.get("latest")
        latest_is_numeric = _set_num(c_l, latest_value)
        if not latest_is_numeric:
            c_l.value = _safe_cell(latest_value)
        else:
            _apply_progress_monetary_format(c_l, display_metric)
        latest_txt = str(c_l.value or "")
        latest_moneyish = _progress_moneyish_text(latest_value if not latest_is_numeric else c_l.value)
        c_l.alignment = Alignment(
            horizontal="right" if (latest_is_numeric or latest_moneyish) else "left",
            vertical="center" if (latest_is_numeric or latest_moneyish or len(latest_txt) <= 18) else "top",
            wrap_text=(not latest_is_numeric and not latest_moneyish and len(latest_txt) > 18),
        )
        status_raw = str(item.get("status") or "").strip().lower()
        status_key = re.sub(r"[\s\-]+", "_", status_raw)
        status_basis = glx_normalize_text(
            " | ".join(
                [
                    str(item.get("latest") or ""),
                    str(item.get("rationale") or ""),
                ]
            )
        ).lower()
        status_class = "open"
        status_disp = "Open"
        if status_key in {"resolved_beat", "actual_beat", "ahead_of_plan", "beat"}:
            status_class = "beat"
            status_disp = "Beat"
        elif status_key in {"resolved_pass", "actual_hit", "hit"}:
            status_class = "hit"
            status_disp = "Hit"
        elif status_key in {"broken", "missed", "resolved_fail", "actual_miss", "miss"}:
            status_class = "missed"
            status_disp = "Missed"
        elif status_key in {"completed", "achieved"}:
            status_class = "completed"
            status_disp = "Completed"
        elif status_key == "on_track" or (status_key == "in_progress" and re.search(r"\bon track\b", status_basis, re.I)):
            status_class = "on_track"
            status_disp = "On track"
        elif status_key == "in_progress":
            status_class = "updated"
            status_disp = "Updated"
        c_s = _ws.cell(row=row_idx, column=4, value=_excel_safe_text_local(status_disp))
        c_s.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        if status_class == "missed":
            c_s.fill = PatternFill("solid", fgColor="FFC7CE")
        elif status_key == "at_risk":
            c_s.fill = PatternFill("solid", fgColor="FFEB9C")
        elif status_class == "beat":
            c_s.fill = PatternFill("solid", fgColor="A9D18E")
        elif status_class == "hit":
            c_s.fill = PatternFill("solid", fgColor="C6EFCE")
        elif status_class == "completed":
            c_s.fill = PatternFill("solid", fgColor="C6EFCE")
        elif status_class == "on_track":
            c_s.fill = PatternFill("solid", fgColor="E2F0D9")
        elif status_class == "updated":
            c_s.fill = PatternFill("solid", fgColor="D9E1F2")
        else:
            c_s.fill = PatternFill("solid", fgColor="D9E1F2")

        rationale_full = _excel_safe_text_local(item.get("rationale") or "")
        gtype = str(item.get("guidance_type") or "").strip().lower()
        fs_q = str(item.get("first_seen_evidence_quarter_end") or item.get("first_seen_quarter_end") or "").strip()
        ls_q = str(item.get("last_seen_evidence_quarter_end") or item.get("last_seen_quarter_end") or "").strip()
        carried_q = str(item.get("carried_to_quarter_end") or "").strip()
        numeric_update_q = bool(item.get("numeric_update_this_quarter"))
        cq = pd.to_datetime(carried_q, errors="coerce")
        lq = pd.to_datetime(ls_q, errors="coerce")
        if gtype in {"run-rate", "ongoing"} and pd.notna(cq) and pd.notna(lq) and pd.Timestamp(cq) > pd.Timestamp(lq) and not numeric_update_q:
            carry_msg = "Carried forward; no new numeric update."
            rationale_full = f"{rationale_full} | {carry_msg}" if rationale_full else carry_msg
        rationale_snip = _excel_safe_text_local(qn_compact_snippet(rationale_full, 240), max_len=240)
        c_r = _ws.cell(row=row_idx, column=5, value=rationale_snip)
        c_r.alignment = Alignment(wrap_text=True, vertical="top")
        c_r.font = Font(size=12, color="000000")

        stated_disp = _q_label(fs_q) if fs_q else ""
        last_seen_disp = _q_label(ls_q) if ls_q else ""
        carried_disp = ""
        if pd.notna(cq):
            carried_disp = _q_label(cq)
        eval_through = _excel_safe_text_local(str(item.get("evaluated_through") or "").strip())
        _ws.cell(row=row_idx, column=6, value=_excel_safe_text_local(stated_disp)).alignment = Alignment(vertical="center", wrap_text=False)
        _ws.cell(row=row_idx, column=7, value=_excel_safe_text_local(last_seen_disp)).alignment = Alignment(vertical="center", wrap_text=False)
        _ws.cell(row=row_idx, column=8, value=_excel_safe_text_local(carried_disp)).alignment = Alignment(vertical="center", wrap_text=False)
        _ws.cell(row=row_idx, column=9, value=eval_through).alignment = Alignment(vertical="center", wrap_text=False)

        qa_sev = str(item.get("qa_severity") or "")
        qa_msg = str(item.get("qa_message") or "")
        ev_row = ev_map_q.get((pid, qd)) or ev_map_pid.get(pid)
        c_e = _ws.cell(row=row_idx, column=10, value="source" if ev_row is not None else "")
        c_e.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        src_txt = ""
        if ev_row is not None:
            _apply_hyperlink_look(c_e, f"#'Promise_Evidence'!A{ev_row}")
        else:
            src_obj = dict(item.get("source") or {})
            src_bits = [
                f"Source: {src_obj.get('source_type') or 'n/a'}",
                f"form={src_obj.get('form') or ''}",
                f"accn={src_obj.get('accn') or ''}",
                f"doc={src_obj.get('doc') or ''}",
                f"section={src_obj.get('section') or ''}",
            ]
            src_txt = _excel_safe_text_local(" | ".join([x for x in src_bits if x]).strip())

        cm = [
            f"Rationale: {rationale_full}" if rationale_full else "",
            f"QA: {qa_sev} {qa_msg}".strip(),
            src_txt,
        ]
        cm_txt = _excel_safe_text_local("\n\n".join([x for x in cm if x]), max_len=32000)
        if cm_txt:
            try:
                _set_cell_comment_local(c_r, cm_txt)
            except Exception:
                pass
        c_width = _ws.column_dimensions["E"].width
        if c_width in (None, 0) or float(c_width) < 40.0:
            c_width = pp_rationale_col_width_default
        row_h = _estimate_wrapped_row_height(
            rationale_snip,
            float(c_width),
            base_height=18.0,
            line_height=14.0,
            min_lines=1,
            max_lines=5,
        )
        rationale_lines = _estimate_wrapped_line_count(
            rationale_snip,
            float(c_width),
            min_lines=1,
            max_lines=5,
        )
        if rationale_lines > 2.2:
            row_h = max(row_h, 50.0 if rationale_lines <= 3.15 else 58.0)
        secondary_h = 0.0
        for cc in range(1, 5):
            cell_txt = str(_ws.cell(row=row_idx, column=cc).value or "")
            if "\n" not in cell_txt and not (cc in {2, 3} and len(cell_txt.strip()) > 14):
                continue
            secondary_h = max(
                secondary_h,
                _estimate_wrapped_row_height(
                    cell_txt,
                    float(_ws.column_dimensions[get_column_letter(cc)].width or 20.0),
                    base_height=18.0,
                    line_height=11.0,
                    min_lines=1,
                    max_lines=3,
                ),
            )
        row_h = max(row_h, secondary_h)
        row_h = min(66.0, row_h)
        _ws.row_dimensions[row_idx].height = row_h

    return _row_writer
