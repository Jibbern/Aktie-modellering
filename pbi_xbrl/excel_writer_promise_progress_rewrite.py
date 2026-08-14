"""Promise_Progress_UI rewrite pass."""
from __future__ import annotations

import math
import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from typing import Any, Dict, Iterable, List, Mapping, MutableMapping, Optional, Sequence, Set, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class PromiseProgressRewriteDeps:
    runtime: MutableMapping[str, Any]


def rewrite_shared_promise_progress_ui_from_blocks(
    deps: PromiseProgressRewriteDeps,
    ws: Any,
    ticker: Any = "",
) -> None:
    """Rewrite legacy stacked progress blocks into the shared Promise dashboard."""
    runtime = deps.runtime
    PROMISE_TIMELINE_HEADERS = runtime["PROMISE_TIMELINE_HEADERS"]
    PROMISE_VISIBLE_MAX_COL = runtime["PROMISE_VISIBLE_MAX_COL"]
    _date_or_none = runtime["_date_or_none"]
    _date_is_missing_or_outside = runtime["_date_is_missing_or_outside"]
    _shared_visible_period_text = runtime["_shared_visible_period_text"]
    _gpre_45z_all_facilities_confirmed = runtime["_gpre_45z_all_facilities_confirmed"]
    _management_credibility_scorecard_rows = runtime["_management_credibility_scorecard_rows"]
    _promise_metric_definition_key = runtime["_promise_metric_definition_key"]
    _promise_progress_label = runtime["_promise_progress_label"]
    _promise_value_looks_like_progress = runtime["_promise_value_looks_like_progress"]
    _remove_empty_promise_revision_blocks = runtime["_remove_empty_promise_revision_blocks"]
    _polish_promise_scorecard_layout = runtime["_polish_promise_scorecard_layout"]
    _finalize_promise_revision_semantics = runtime["_finalize_promise_revision_semantics"]
    _apply_source_backed_promise_mapping_overrides = runtime["_apply_source_backed_promise_mapping_overrides"]
    _apply_promise_grid_style = runtime["_apply_promise_grid_style"]

    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    ticker_txt = str(ticker or "").strip().upper()
    if ticker_txt == "ANF":
        return
    old_rows: List[Dict[str, Any]] = []
    current_asof = ""
    current_asof_date: Optional[date] = None
    current_block_label = ""
    rr = 1
    while rr <= int(ws.max_row or 0):
        a_val = str(ws.cell(rr, 1).value or "").strip()
        m = re.match(r"Promise progress \(As of (\d{4}-\d{2}-\d{2})\)", a_val)
        if m:
            current_asof = m.group(1)
            try:
                current_asof_date = pd.Timestamp(current_asof).date()
            except Exception:
                current_asof_date = None
            if isinstance(current_asof_date, date):
                qn = ((int(current_asof_date.month) - 1) // 3) + 1
                current_block_label = f"{int(current_asof_date.year)}-Q{qn}"
            else:
                current_block_label = current_asof
            rr += 2  # skip legacy header row
            continue
        if current_asof and a_val and a_val.lower() != "metric":
            metric = _shared_visible_period_text(a_val)
            metric_low = metric.lower()
            is_actual_only_metric = (
                metric_low.endswith(" actual")
                or (re.search(r"\bactual\b", metric_low) and "guidance" not in metric_low and "target" not in metric_low)
            )
            if metric.lower() != "no high-signal items." and not is_actual_only_metric:
                old_rows.append(
                    {
                        "metric": metric,
                        "target": _shared_visible_period_text(str(ws.cell(rr, 2).value or "")),
                        "latest": _shared_visible_period_text(str(ws.cell(rr, 3).value or "")),
                        "status": _shared_visible_period_text(str(ws.cell(rr, 4).value or "")),
                        "note": _shared_visible_period_text(str(ws.cell(rr, 5).value or "")),
                        "stated": _shared_visible_period_text(str(ws.cell(rr, 6).value or "")) or current_block_label,
                        "evaluated_through": _shared_visible_period_text(str(ws.cell(rr, 9).value or "")),
                        "source_date": current_asof,
                        "asof_date": current_asof_date,
                        "block_label": current_block_label,
                    }
                )
        rr += 1
    if not old_rows:
        return

    def _canonical_promise_status(value_in: Any, *, metric_hint: Any = "", horizon_hint: Any = "") -> str:
        raw = _shared_visible_period_text(str(value_in or "")).strip()
        low = raw.lower()
        if low in {"", "nan", "none", "null"}:
            return ""
        if low in {"completed", "complete", "delivered", "achieved"}:
            return "Completed"
        if low in {"hit"}:
            return "Hit"
        if low in {"met"}:
            return "Completed"
        if low in {"beat"}:
            return "Beat"
        if low in {"miss", "missed", "fail", "failed"}:
            return "Missed"
        if low in {"on track", "on_track", "partial", "met-ish", "met ish"}:
            return "On track"
        if low in {"updated", "raised", "lowered", "maintained"}:
            return "On track"
        if low in {"open", "not yet measurable", "not yet realized", "not yet assessed"}:
            return "Open"
        if low in {"mixed", "met/miss"}:
            return "Mixed"
        if low in {"basis-dependent", "basis dependent"}:
            return "Basis-dependent"
        if low in {"n/a", "na", "not assessed", "not assessable"}:
            return "Not assessed"
        return raw

    def _format_promise_value_text(value_in: Any, metric_hint: Any = "") -> str:
        txt = _shared_visible_period_text(str(value_in or "")).strip()
        if not txt:
            return ""
        txt = re.sub(
            r"\ball\s+8\s+plants\s+qualified\s+for\s+45z\s+tax\s+credits\b",
            "All 8 qualified",
            txt,
            flags=re.I,
        )
        txt = re.sub(
            r"\ball\s+eight\s+operating\s+plants\s+qualified/expected\s+to\s+qualify\s+for\s+45z\s+tax\s+credits\s+in\s+2026\b",
            "All 8 qualified",
            txt,
            flags=re.I,
        )
        txt = re.sub(
            r"\ball\s+8\s+plants\s+qualified/expected\s+to\s+qualify(?:\s+in\s+2026)?\b",
            "All 8 qualified",
            txt,
            flags=re.I,
        )
        txt = re.sub(
            r"\badvantage\s+nebraska\s+fully\s+operational(?:\s+and\s+sequestering\s+co2\s+in\s+wyoming)?\b",
            "AN operational",
            txt,
            flags=re.I,
        )
        txt = re.sub(
            r"\bstrategic\s+review\s+phase\s+2\s+remains\s+on\s+track\s+by\s+end\s+of\s+2026-Q2\b",
            "Phase 2 on track",
            txt,
            flags=re.I,
        )
        txt = re.sub(
            r"\bstrategic\s+review\s+phase\s+2\s+was\s+initiated\s+in\s+(?:Q2\s+2026|2026-Q2)\b",
            "Phase 2 initiated",
            txt,
            flags=re.I,
        )
        if re.fullmatch(r"-?\d+(?:\.\d+)?", txt):
            try:
                val = float(txt)
            except Exception:
                return txt
            metric_low = str(metric_hint or "").lower()
            if "eps" in metric_low and abs(val) < 100:
                return f"${val:,.2f}"
            abs_val = abs(val)
            if abs_val >= 1_000_000_000:
                return f"${val / 1_000_000_000:,.2f}bn"
            if abs_val >= 1_000_000:
                return f"${val / 1_000_000:,.1f}m"
            if abs_val >= 1_000:
                return f"${val:,.0f}"
        return txt

    def _promise_actual_lookup_from_workbook() -> Tuple[Dict[str, Dict[str, float]], Dict[str, Dict[str, float]], Dict[int, Dict[str, float]], Dict[int, date]]:
        wb = getattr(ws, "parent", None)
        by_period: Dict[str, Dict[str, float]] = {}
        by_ytd_period: Dict[str, Dict[str, float]] = {}
        by_year: Dict[int, Dict[str, float]] = {}
        year_end_dates: Dict[int, date] = {}
        history_period_labels_by_date: Dict[date, Set[str]] = {}
        if wb is None:
            return by_period, by_ytd_period, by_year, year_end_dates

        def _norm_header(value: Any) -> str:
            return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")

        def _num(value: Any) -> Optional[float]:
            out = pd.to_numeric(value, errors="coerce")
            if pd.isna(out):
                return None
            val = float(out)
            return val if math.isfinite(val) else None

        def _period_labels(qd: date, row_map: Mapping[str, Any]) -> List[str]:
            labels = {str(qd)}
            fiscal_labels: Set[str] = set()
            fiscal_label = str(row_map.get("fiscal_label") or "").strip()
            if fiscal_label:
                fiscal_labels.add(fiscal_label)
            fy = _num(row_map.get("fiscal_year"))
            fq = _num(row_map.get("fiscal_quarter"))
            if fy is not None and fq is not None and 1 <= int(fq) <= 4:
                fiscal_labels.add(f"{int(fy)}-Q{int(fq)}")
            if fiscal_labels:
                labels.update(fiscal_labels)
            else:
                labels.add(f"{qd.year}-Q{((qd.month - 1) // 3) + 1}")
            return [label for label in labels if label]

        def _year_for_row(qd: date, row_map: Mapping[str, Any]) -> int:
            fy = _num(row_map.get("fiscal_year"))
            return int(fy) if fy is not None else int(qd.year)

        def _add_period(labels: Iterable[str], key: str, value: Optional[float]) -> None:
            if value is None:
                return
            for label in labels:
                by_period.setdefault(label, {})[key] = value

        def _add_year(year: int, key: str, value: Optional[float]) -> None:
            if value is None:
                return
            by_year.setdefault(int(year), {})[key] = by_year.setdefault(int(year), {}).get(key, 0.0) + value

        def _labels_year_or_calendar(qd: date, fiscal_year: Optional[float], fiscal_quarter: Optional[float]) -> Tuple[Set[str], int]:
            labels: Set[str] = {str(qd)}
            if fiscal_year is not None and fiscal_quarter is not None and 1 <= int(fiscal_quarter) <= 4:
                labels.add(f"{int(fiscal_year)}-Q{int(fiscal_quarter)}")
                return labels, int(fiscal_year)
            hist_labels = history_period_labels_by_date.get(qd)
            if hist_labels:
                labels.update(hist_labels)
                year_match = next((re.match(r"^(20\d{2})-Q[1-4]$", label) for label in sorted(hist_labels)), None)
                if year_match:
                    return labels, int(year_match.group(1))
            labels.add(f"{qd.year}-Q{((qd.month - 1) // 3) + 1}")
            return labels, int(qd.year)

        if "History_Q" in getattr(wb, "sheetnames", []):
            hist_ws = wb["History_Q"]
            headers = {
                _norm_header(hist_ws.cell(1, cc).value): cc
                for cc in range(1, int(hist_ws.max_column or 0) + 1)
            }
            q_col = headers.get("quarter")
            if q_col:
                for rr in range(2, int(hist_ws.max_row or 0) + 1):
                    qd = _date_or_none(hist_ws.cell(rr, q_col).value)
                    if qd is None:
                        continue
                    row_map = {
                        name: hist_ws.cell(rr, cc).value
                        for name, cc in headers.items()
                    }
                    labels = _period_labels(qd, row_map)
                    history_period_labels_by_date.setdefault(qd, set()).update(
                        label for label in labels if re.fullmatch(r"20\d{2}-Q[1-4]", str(label))
                    )
                    year = _year_for_row(qd, row_map)
                    year_end_dates[year] = max(year_end_dates.get(year, date.min), qd)
                    revenue = _num(row_map.get("revenue"))
                    op_income = _num(row_map.get("op_income"))
                    cfo = _num(row_map.get("cfo"))
                    capex = _num(row_map.get("capex"))
                    eps = _num(row_map.get("eps_diluted"))
                    shares = _num(row_map.get("shares_diluted"))
                    buybacks = _num(row_map.get("buybacks_cash"))
                    operating_margin = None
                    if revenue and op_income is not None:
                        operating_margin = op_income / revenue
                    fcf = (cfo - capex) if cfo is not None and capex is not None else None
                    for key, val in (
                        ("revenue", revenue),
                        ("op_income", op_income),
                        ("operating_margin", operating_margin),
                        ("fcf", fcf),
                        ("capex", capex),
                        ("eps", eps),
                        ("shares", shares),
                        ("buybacks", buybacks),
                    ):
                        _add_period(labels, key, val)
                        _add_year(year, key, val)

        if "Adjusted_Metrics" in getattr(wb, "sheetnames", []):
            adj_ws = wb["Adjusted_Metrics"]
            headers = {
                _norm_header(adj_ws.cell(1, cc).value): cc
                for cc in range(1, int(adj_ws.max_column or 0) + 1)
            }
            q_col = headers.get("quarter")
            if q_col:
                for rr in range(2, int(adj_ws.max_row or 0) + 1):
                    qd = _date_or_none(adj_ws.cell(rr, q_col).value)
                    if qd is None:
                        continue
                    period_type = str(adj_ws.cell(rr, headers.get("period_type", 0)).value or "").strip().lower() if headers.get("period_type") else ""
                    if period_type and "annual" in period_type:
                        continue
                    row_map = {name: adj_ws.cell(rr, cc).value for name, cc in headers.items()}
                    fiscal_year = _num(row_map.get("fiscal_year"))
                    fiscal_quarter = _num(row_map.get("fiscal_quarter"))
                    labels, year = _labels_year_or_calendar(qd, fiscal_year, fiscal_quarter)
                    year_end_dates[year] = max(year_end_dates.get(year, date.min), qd)
                    for key, source_name in (
                        ("adj_ebit", "adj_ebit"),
                        ("adj_ebitda", "adj_ebitda"),
                        ("adj_eps", "adj_eps"),
                        ("adj_fcf", "adj_fcf"),
                    ):
                        val = _num(row_map.get(source_name))
                        _add_period(labels, key, val)
                        _add_year(year, key, val)

        def _build_ytd_periods() -> None:
            additive_keys = {
                "revenue",
                "op_income",
                "fcf",
                "capex",
                "buybacks",
                "adj_ebit",
                "adj_ebitda",
                "adj_fcf",
            }
            quarter_labels = sorted(
                {
                    str(label)
                    for label in by_period
                    if re.fullmatch(r"20\d{2}-Q[1-4]", str(label))
                }
            )
            for label in quarter_labels:
                m = re.fullmatch(r"(20\d{2})-Q([1-4])", label)
                if not m:
                    continue
                year = int(m.group(1))
                qtr = int(m.group(2))
                labels = [f"{year}-Q{idx}" for idx in range(1, qtr + 1)]
                ytd_vals: Dict[str, float] = {}
                for key in additive_keys:
                    vals: List[float] = []
                    complete = True
                    for period_label in labels:
                        period_vals = by_period.get(period_label, {})
                        if key not in period_vals:
                            complete = False
                            break
                        vals.append(float(period_vals[key]))
                    if complete and vals:
                        ytd_vals[key] = sum(vals)
                if "revenue" in ytd_vals and "op_income" in ytd_vals and ytd_vals["revenue"]:
                    ytd_vals["operating_margin"] = ytd_vals["op_income"] / ytd_vals["revenue"]
                if ytd_vals:
                    by_ytd_period[label] = ytd_vals

        _build_ytd_periods()
        return by_period, by_ytd_period, by_year, year_end_dates

    actuals_by_period, actuals_ytd_by_period, actuals_by_year, year_end_dates = _promise_actual_lookup_from_workbook()

    def _promise_metric_actual_key(metric_in: Any) -> str:
        metric_low = str(metric_in or "").strip().lower()
        if "adjusted ebitda" in metric_low or "adj ebitda" in metric_low:
            return "adj_ebitda"
        if "adjusted ebit" in metric_low or "adj ebit" in metric_low:
            return "adj_ebit"
        if "fcf" in metric_low or "free cash" in metric_low:
            return "adj_fcf" if "adjusted" in metric_low or "adj" in metric_low else "fcf"
        if "capex" in metric_low or "capital expenditure" in metric_low:
            return "capex"
        if "operating margin" in metric_low or "op margin" in metric_low:
            return "operating_margin"
        if "eps" in metric_low:
            return "adj_eps" if "adjusted" in metric_low or "adj" in metric_low else "eps"
        if "share" in metric_low and "repurchase" not in metric_low and "buyback" not in metric_low:
            return "shares"
        if "repurchase" in metric_low or "buyback" in metric_low:
            return "buybacks"
        if "revenue" in metric_low or "sales" in metric_low:
            return "revenue"
        return ""

    def _format_lookup_actual_value(metric_in: Any, key: str, value: Any) -> str:
        val = pd.to_numeric(value, errors="coerce")
        if pd.isna(val):
            return ""
        num = float(val)
        if key == "operating_margin":
            return f"{num * 100:.1f}%"
        if key in {"adj_eps", "eps"}:
            return f"${num:,.2f}"
        if key == "shares":
            return f"{num / 1_000_000:,.1f}m" if abs(num) > 1_000_000 else f"{num:,.1f}m"
        return _format_promise_value_text(str(num), metric_in)

    def _actual_for_stated_quarter(row: Mapping[str, Any]) -> str:
        key = _promise_metric_actual_key(row.get("metric"))
        if not key:
            return ""
        stated_txt = str(row.get("stated") or row.get("block_label") or "").strip()
        labels = [stated_txt]
        m = re.search(r"\b(20\d{2})-Q([1-4])\b", stated_txt, flags=re.I)
        if m:
            labels.append(f"{int(m.group(1))}-Q{int(m.group(2))}")
        for label in labels:
            period_vals = actuals_by_period.get(label, {})
            if key in period_vals:
                return _format_lookup_actual_value(row.get("metric"), key, period_vals[key])
            if key == "eps" and "adj_eps" in period_vals:
                return _format_lookup_actual_value(row.get("metric"), "adj_eps", period_vals["adj_eps"])
            if key == "adj_fcf" and "fcf" in period_vals:
                return _format_lookup_actual_value(row.get("metric"), "fcf", period_vals["fcf"])
        return ""

    def _ytd_for_stated_quarter(row: Mapping[str, Any]) -> str:
        key = _promise_metric_actual_key(row.get("metric"))
        if not key or key in {"eps", "adj_eps", "shares"}:
            return ""
        stated_txt = str(row.get("stated") or row.get("block_label") or "").strip()
        labels = [stated_txt]
        m = re.search(r"\b(20\d{2})-Q([1-4])\b", stated_txt, flags=re.I)
        if m:
            labels.append(f"{int(m.group(1))}-Q{int(m.group(2))}")
        for label in labels:
            ytd_vals = actuals_ytd_by_period.get(label, {})
            if key in ytd_vals:
                return _format_lookup_actual_value(row.get("metric"), key, ytd_vals[key])
            if key == "adj_fcf" and "fcf" in ytd_vals:
                return _format_lookup_actual_value(row.get("metric"), "fcf", ytd_vals["fcf"])
        return ""

    def _annual_actual_for_metric(year: int, metric_in: Any) -> str:
        key = _promise_metric_actual_key(metric_in)
        if not key:
            return ""
        annual_vals = actuals_by_year.get(int(year), {})
        if key in annual_vals:
            return _format_lookup_actual_value(metric_in, key, annual_vals[key])
        if key == "eps" and "adj_eps" in annual_vals:
            return _format_lookup_actual_value(metric_in, "adj_eps", annual_vals["adj_eps"])
        if key == "adj_fcf" and "fcf" in annual_vals:
            return _format_lookup_actual_value(metric_in, "fcf", annual_vals["fcf"])
        return ""

    def _numbers_for_status(txt: Any) -> List[float]:
        values: List[float] = []
        for match in re.findall(r"(?<![A-Za-z])\d+(?:,\d{3})*(?:\.\d+)?", str(txt or "")):
            try:
                values.append(float(match.replace(",", "")))
            except Exception:
                continue
        return values

    def _status_from_guidance_actual(metric_in: Any, guide_in: Any, actual_in: Any) -> str:
        metric_low = str(metric_in or "").lower()
        guide_nums = _numbers_for_status(guide_in)
        actual_nums = _numbers_for_status(actual_in)
        if not actual_nums:
            return ""
        if not guide_nums:
            return "Hit"
        actual_val = actual_nums[0]
        low = min(guide_nums)
        high = max(guide_nums)
        if len(guide_nums) == 1:
            low = high = guide_nums[0]
        tolerance = max(abs(high) * 0.001, 0.001)
        if "capex" in metric_low or "capital expenditure" in metric_low:
            return "Hit" if low - tolerance <= actual_val <= high + tolerance else "Mixed"
        if "share" in metric_low and "repurchase" not in metric_low and "buyback" not in metric_low:
            return "Hit" if low - tolerance <= actual_val <= high + tolerance else "Mixed"
        if actual_val > high + tolerance:
            return "Beat"
        if actual_val < low - tolerance:
            return "Missed"
        return "Hit"

    def _horizon_from_row(row: Dict[str, Any]) -> str:
        blob = " ".join(str(row.get(k) or "") for k in ("note", "target", "metric"))
        blob = _shared_visible_period_text(blob)
        blob_low = blob.lower()
        if "cost savings" in blob_low and "annualized" in blob_low:
            return "Annualized program"
        if "interest expense" in blob_low and re.search(r"\b2026\b", blob_low):
            return "2026 year"
        if ticker_txt == "GPRE" and "45z" in blob_low and (
            "facility qualification" in blob_low
            or "qualified" in blob_low
            or "qualify" in blob_low
        ) and re.search(r"\b2026\b", blob_low):
            return "2026-Q1"
        if ticker_txt == "GPRE" and "45z" in blob_low and re.search(r"\b2026\b", blob_low) and (
            "remaining facilities" in blob_low
            or "advantage nebraska" in blob_low
            or "$200m-$225m" in blob_low
            or "$188" in blob_low
            or "188.0m" in blob_low
            or "200m-$225m" in blob_low
        ):
            return "2026 year"
        m = re.search(r"\b(20\d{2})\s+year\b", blob, re.I)
        if m:
            return f"{m.group(1)} year"
        m = re.search(r"\b(20\d{2}-Q[1-4])\b", blob, re.I)
        if m:
            return m.group(1)
        return str(row.get("stated") or row.get("block_label") or "")

    def _canonical_timeline_metric(metric_in: Any, horizon_in: Any, row_in: Mapping[str, Any]) -> str:
        metric_txt = str(metric_in or "").strip()
        metric_low = metric_txt.lower()
        blob_low = " ".join(str(row_in.get(k) or "") for k in ("metric", "target", "latest", "note")).lower()
        if ticker_txt == "GPRE" and "45z" in blob_low and "2026 year" == str(horizon_in or "").strip().lower():
            if "ebitda" in blob_low or "adjusted ebitda" in blob_low or "$188" in blob_low or "$200" in blob_low:
                return "2026 year 45Z EBITDA guidance"
        if ticker_txt == "GPRE" and ("facility qualification" in metric_low or ("45z" in blob_low and ("qualified" in blob_low or "qualify" in blob_low))):
            return "45Z facility qualification"
        return metric_txt

    def _horizon_end_from_label(label: Any) -> Optional[date]:
        txt = str(label or "").strip()
        m = re.fullmatch(r"(20\d{2})\s+year", txt, flags=re.I)
        if m:
            return date(int(m.group(1)), 12, 31)
        m = re.fullmatch(r"(20\d{2})-Q([1-4])", txt, flags=re.I)
        if m:
            yy = int(m.group(1))
            qq = int(m.group(2))
            if qq == 1:
                return date(yy, 3, 31)
            if qq == 2:
                return date(yy, 6, 30)
            if qq == 3:
                return date(yy, 9, 30)
            return date(yy, 12, 31)
        return None

    def _has_measurable_actual_text(value_in: Any) -> bool:
        txt = str(value_in or "").strip()
        if not txt:
            return False
        if re.search(r"\b(not yet|not measurable|not assessed|open|expected|expected to|pending)\b", txt, flags=re.I):
            return False
        return bool(re.search(r"\d|completed|fully operational|qualified|actual|hit|beat|miss", txt, flags=re.I))

    def _is_partial_tracking_text(value_in: Any) -> bool:
        txt = str(value_in or "").strip()
        if not txt:
            return False
        return bool(
            re.search(
                r"\b(Q[1-3]|first quarter|second quarter|third quarter|YTD|year-to-date|run[- ]rate|partial|latest tracking|latest)\b",
                txt,
                flags=re.I,
            )
        )

    def _partial_tracking_conflicts_with_horizon(value_in: Any, horizon_txt: Any) -> bool:
        txt = str(value_in or "").strip()
        if not txt or not _is_partial_tracking_text(txt):
            return False
        horizon = str(horizon_txt or "").strip()
        m_h = re.fullmatch(r"(20\d{2})-Q([1-4])", horizon, flags=re.I)
        if m_h:
            h_year = int(m_h.group(1))
            h_q = int(m_h.group(2))
            q_mentions = {int(q) for q in re.findall(r"\bQ([1-4])\b", txt, flags=re.I)}
            if q_mentions and h_q not in q_mentions:
                return True
            y_mentions = {int(y) for y in re.findall(r"\b(20\d{2})\b", txt)}
            if y_mentions and h_year not in y_mentions:
                return True
        return False

    def _row_evaluation_date(row: Mapping[str, Any]) -> date:
        for key in ("evaluated_through", "source_date"):
            try:
                txt = str(row.get(key) or "").strip()
                if txt:
                    return pd.Timestamp(txt).date()
            except Exception:
                continue
        return date.min

    def _promise_actual_for_horizon(row: Mapping[str, Any], horizon_txt: Any, status_txt: Any) -> str:
        latest_txt = str(row.get("latest") or "").strip()
        if not latest_txt or not _has_measurable_actual_text(latest_txt):
            return ""
        if _partial_tracking_conflicts_with_horizon(latest_txt, horizon_txt):
            return ""
        status_low = str(_canonical_promise_status(status_txt, metric_hint=row.get("metric"), horizon_hint=horizon_txt)).strip().lower()
        complete_statuses = {"completed", "beat", "hit", "missed", "basis-dependent", "mixed"}
        horizon_end = _horizon_end_from_label(horizon_txt)
        evaluation_dt = _row_evaluation_date(row)
        annual_match = re.fullmatch(r"(20\d{2})\s+year", str(horizon_txt or "").strip(), flags=re.I)
        if annual_match:
            if evaluation_dt < (horizon_end or date(int(annual_match.group(1)), 12, 31)):
                return ""
            if _is_partial_tracking_text(latest_txt):
                return ""
            return _format_promise_value_text(latest_txt, row.get("metric")) if status_low in complete_statuses else ""
        if horizon_end is not None and evaluation_dt < horizon_end:
            return ""
        if status_low in complete_statuses or re.search(r"\bactual\b", str(row.get("metric") or ""), flags=re.I):
            return _format_promise_value_text(latest_txt, row.get("metric"))
        return ""

    def _promise_lifecycle_id(metric_in: Any, horizon_in: Any) -> str:
        metric_slug = re.sub(r"[^a-z0-9]+", "_", str(metric_in or "").strip().lower()).strip("_")
        horizon_slug = re.sub(r"[^a-z0-9]+", "_", str(horizon_in or "").strip().lower()).strip("_")
        if not metric_slug:
            return ""
        if metric_slug in {"cost_savings", "cost_savings_target", "cost_reduction", "cost_reduction_target"}:
            return "guidance:cost_savings:ANNUALIZED_PROGRAM"
        return f"guidance:{metric_slug}:{horizon_slug}" if horizon_slug else f"guidance:{metric_slug}"

    rows_sorted = sorted(
        old_rows,
        key=lambda row: row.get("asof_date") or date.min,
        reverse=True,
    )
    row_horizons: Dict[int, str] = {idx: _horizon_from_row(row) for idx, row in enumerate(rows_sorted)}
    older_by_metric: Dict[int, Dict[str, Any]] = {}
    for idx, row in enumerate(rows_sorted):
        row_horizon = row_horizons.get(idx, "")
        for j, cand in enumerate(rows_sorted):
            if j <= idx:
                continue
            if (
                str(cand.get("metric") or "").lower() == str(row.get("metric") or "").lower()
                and row_horizons.get(j, "") == row_horizon
            ):
                older_by_metric[idx] = cand
                break

    def _stated_quarter(label: Any) -> Tuple[Optional[int], Optional[int]]:
        m = re.search(r"\b(20\d{2})-Q([1-4])\b", str(label or ""), flags=re.I)
        if not m:
            return None, None
        return int(m.group(1)), int(m.group(2))

    def _annual_year_from_horizon(label: Any) -> Optional[int]:
        m = re.fullmatch(r"(20\d{2})\s+year", str(label or "").strip(), flags=re.I)
        return int(m.group(1)) if m else None

    timeline_rows: List[Dict[str, str]] = []
    for idx, row in enumerate(rows_sorted):
        older = older_by_metric.get(idx)
        prev = str((older or {}).get("target") or "").strip()
        new = str(row.get("target") or "").strip()
        if not prev:
            change = "Initial"
        elif prev == new:
            change = "Maintained"
        else:
            change = "Updated"
        horizon_txt = row_horizons.get(idx, "") or _horizon_from_row(row)
        status_txt = _canonical_promise_status(row.get("status"), metric_hint=row.get("metric"), horizon_hint=horizon_txt)
        if ticker_txt == "GPRE":
            metric_low = str(row.get("metric") or "").strip().lower()
            stated_txt = str(row.get("stated") or row.get("block_label") or "").strip()
            timing_blob = " ".join(
                str(row.get(k) or "")
                for k in ("latest", "note", "target", "source_date")
            )
            if (
                metric_low == "cost savings target"
                and stated_txt == "2025-Q1"
                and re.search(r"\b2025-Q2\b|\bQ2\s+2025\b", timing_blob, flags=re.I)
            ):
                continue
        annual_match = re.fullmatch(r"(20\d{2})\s+year", str(horizon_txt or "").strip(), flags=re.I)
        if annual_match and str(status_txt).strip().lower() in {"completed", "beat", "hit", "missed"}:
            latest_txt = str(row.get("latest") or "").strip()
            try:
                horizon_year = int(annual_match.group(1))
            except Exception:
                horizon_year = 0
            evaluation_dt = _row_evaluation_date(row)
            if horizon_year and evaluation_dt < date(horizon_year, 12, 31):
                note_txt = str(row.get("note") or "").strip()
                partial_note = f"Partial/latest tracking only; {horizon_year} annual horizon is not complete."
                if latest_txt and partial_note.lower() not in note_txt.lower():
                    row["note"] = f"{note_txt} {partial_note}".strip()
                status_txt = "On track" if latest_txt else "Open"
        actual_for_horizon = _promise_actual_for_horizon(row, horizon_txt, status_txt)
        progress_for_horizon = ""
        if (
            not actual_for_horizon
            and str(status_txt).strip().lower() in {"completed", "beat", "hit", "missed", "mixed", "basis-dependent"}
            and _partial_tracking_conflicts_with_horizon(row.get("latest"), horizon_txt)
        ):
            status_txt = "On track" if str(row.get("latest") or "").strip() else "Open"
        if annual_match:
            try:
                horizon_year = int(annual_match.group(1))
            except Exception:
                horizon_year = 0
            stated_year, stated_q = _stated_quarter(row.get("stated") or row.get("block_label"))
            if horizon_year and stated_year == horizon_year and stated_q in {1, 2, 3}:
                stated_actual = _actual_for_stated_quarter(row)
                if stated_actual:
                    actual_for_horizon = stated_actual
                    ytd_actual = _ytd_for_stated_quarter(row)
                    if ytd_actual and not progress_for_horizon:
                        progress_for_horizon = f"YTD: {ytd_actual}"
                    if status_txt in {"Completed", "Beat", "Hit", "Missed", "Mixed", "Basis-dependent"}:
                        status_txt = "On track"
                elif actual_for_horizon:
                    # Legacy carried-forward promise blocks may include a
                    # final annual result on an earlier stated-in quarter once
                    # the horizon has been evaluated. Keep that final result
                    # in the Q4/final row and annual progression table, but do
                    # not display it as the Q1/Q2/Q3 quarter actual.
                    actual_for_horizon = ""
                    if status_txt in {"Completed", "Beat", "Hit", "Missed", "Mixed", "Basis-dependent"}:
                        status_txt = "On track"
        stated_out = str(row.get("stated") or row.get("block_label") or "")
        source_out = str(row.get("source_date") or "")
        note_out = str(row.get("note") or "")
        metric_out = _canonical_timeline_metric(row.get("metric"), horizon_txt, row)
        if ticker_txt == "GPRE" and metric_out == "45Z facility qualification" and horizon_txt == "2026-Q1":
            facility_blob = " | ".join(
                str(x or "")
                for x in (
                    prev,
                    new,
                    metric_out,
                    actual_for_horizon,
                    progress_for_horizon,
                    row.get("latest"),
                    note_out,
                    row.get("source_note"),
                )
            )
            if _gpre_45z_all_facilities_confirmed(facility_blob):
                actual_for_horizon = "All 8 qualified/operational"
                progress_for_horizon = ""
                status_txt = "Completed"
                stated_out = "2026-Q1"
                if _date_is_missing_or_outside(source_out, date(2026, 3, 31), date(2026, 12, 31)):
                    source_out = "2026-03-31"
                if not note_out:
                    note_out = "Conference metadata says all plants qualify from Jan. 1; Advantage Nebraska operational."
            elif stated_out == "2026-Q1":
                progress_for_horizon = "3 of 8 qualified"
                actual_for_horizon = ""
                status_txt = "On track"
            elif status_txt in {"Completed", "Hit", "Beat"}:
                status_txt = "On track"
        if _promise_value_looks_like_progress(actual_for_horizon, metric=metric_out):
            progress_for_horizon = progress_for_horizon or _promise_progress_label(
                actual_for_horizon,
                metric=metric_out,
                stated=stated_out,
            )
            actual_for_horizon = ""
            if status_txt in {"Completed", "Beat", "Hit", "Missed"}:
                status_txt = "On track"
        latest_progress = str(row.get("latest") or "").strip()
        if (
            not progress_for_horizon
            and not actual_for_horizon
            and _promise_value_looks_like_progress(latest_progress, metric=metric_out)
        ):
            progress_for_horizon = _promise_progress_label(
                latest_progress,
                metric=metric_out,
                stated=stated_out,
            )
            if status_txt in {"Completed", "Beat", "Hit", "Missed"}:
                status_txt = "On track"
        # Do not turn structural blank rows into visible revision rows simply
        # because the surrounding block supplies stated/source/horizon labels.
        if not any(str(value or "").strip() for value in (prev, new, actual_for_horizon, progress_for_horizon, status_txt, note_out)):
            continue
        timeline_rows.append(
            {
                "Metric": metric_out,
                "Previous guide": _format_promise_value_text(prev, row.get("metric")),
                "New/current guide": _format_promise_value_text(new, row.get("metric")),
                "Change type": change,
                "Actual": actual_for_horizon,
                "Progress / run-rate": progress_for_horizon,
                "Status": status_txt,
                "Horizon": horizon_txt,
                "Stated in": stated_out,
                "Source date": source_out,
                "Source / note": note_out,
            }
        )

    def _quarter_sort_score(label: Any) -> int:
        year, qtr = _stated_quarter(label)
        return year * 10 + qtr if year and qtr else 0

    def _source_date_value(rec: Mapping[str, Any]) -> date:
        try:
            return pd.Timestamp(str(rec.get("Source date") or "")).date()
        except Exception:
            return date.min

    def _display_block_for_timeline_row(rec: Mapping[str, Any]) -> str:
        horizon_txt = str(rec.get("Horizon") or "").strip()
        stated_txt = str(rec.get("Stated in") or "").strip()
        annual_year = _annual_year_from_horizon(horizon_txt)
        stated_year, stated_q = _stated_quarter(stated_txt)
        if annual_year is not None:
            if "pre-release" in stated_txt.lower():
                return f"{annual_year}-Q4 pre-release update"
            if stated_year == annual_year and stated_q:
                return f"{annual_year}-Q{stated_q}"
            if stated_year is not None and stated_year < annual_year:
                return f"{annual_year}-Q1"
            if stated_year is not None and stated_year > annual_year:
                return f"{annual_year}-Q4"
            return f"{annual_year}-Q1"
        horizon_year, horizon_q = _stated_quarter(horizon_txt)
        if horizon_year and horizon_q:
            return f"{horizon_year}-Q{horizon_q}"
        return stated_txt

    def _change_type_from_prev_new(prev_in: Any, new_in: Any, fallback: Any = "") -> str:
        prev_txt = str(prev_in or "").strip()
        new_txt = str(new_in or "").strip()
        if not prev_txt:
            return "Initial"
        if prev_txt == new_txt:
            return "Maintained"
        fallback_txt = str(fallback or "").strip()
        if fallback_txt and fallback_txt.lower() not in {"initial", "maintained"}:
            return fallback_txt
        return "Updated"

    def _normalize_timeline_rows(rows_in: List[Dict[str, str]]) -> List[Dict[str, str]]:
        if not rows_in:
            return []
        rows = [dict(row) for row in rows_in]
        for row in rows:
            row["Display block"] = _display_block_for_timeline_row(row)
            if (
                ticker_txt == "GPRE"
                and str(row.get("Metric") or "").strip() == "2026 year 45Z EBITDA guidance"
                and str(row.get("Horizon") or "").strip() == "2026 year"
                and re.search(r"\$?\s*200m\s*-\s*\$?\s*225m|\$?200m-\$?225m", str(row.get("New/current guide") or ""), flags=re.I)
            ):
                row["Previous guide"] = "$188m"
                row["Change type"] = "Updated"
            if (
                ticker_txt == "GPRE"
                and str(row.get("Metric") or "").strip() == "45Z facility qualification"
                and str(row.get("Horizon") or "").strip() == "2026-Q1"
            ):
                if not str(row.get("Previous guide") or "").strip():
                    row["Previous guide"] = str(row.get("New/current guide") or "").strip()
                if _gpre_45z_all_facilities_confirmed(
                    row.get("Previous guide"),
                    row.get("New/current guide"),
                    row.get("Metric"),
                    row.get("Actual"),
                    row.get("Progress / run-rate"),
                    row.get("Source / note"),
                ):
                    row["Actual"] = "All 8 qualified/operational"
                    row["Progress / run-rate"] = ""
                    row["Status"] = "Completed"
                    if not str(row.get("Source / note") or "").strip():
                        row["Source / note"] = "Conference metadata says all plants qualify from Jan. 1; Advantage Nebraska operational."
                else:
                    row["Actual"] = ""
                    row["Progress / run-rate"] = "3 of 8 qualified"
                    row["Status"] = "On track"
                row["Change type"] = _change_type_from_prev_new(
                    row.get("Previous guide"),
                    row.get("New/current guide"),
                    "Maintained",
                )
                # The underlying guide can be stated earlier, but this display
                # row is the 2026-Q1 progress read for the horizon.
                row["Stated in"] = "2026-Q1"
                if _date_is_missing_or_outside(row.get("Source date"), date(2026, 3, 31), date(2026, 12, 31)):
                    row["Source date"] = "2026-03-31"
        # Fill missing Previous guide from the immediately prior source for
        # the same metric/horizon. This keeps the block based on the horizon
        # while preserving the source quarter in Stated in.
        grouped: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
        for row in rows:
            key = (_promise_metric_definition_key(row.get("Metric")), str(row.get("Horizon") or "").strip().lower())
            grouped.setdefault(key, []).append(row)
        for group_rows in grouped.values():
            ordered = sorted(
                group_rows,
                key=lambda rec: (_source_date_value(rec), _quarter_sort_score(rec.get("Stated in")), _quarter_sort_score(rec.get("Display block"))),
            )
            last_guide = ""
            for rec in ordered:
                if not str(rec.get("Previous guide") or "").strip() and last_guide:
                    rec["Previous guide"] = last_guide
                rec["Change type"] = _change_type_from_prev_new(rec.get("Previous guide"), rec.get("New/current guide"), rec.get("Change type"))
                new_txt = str(rec.get("New/current guide") or "").strip()
                if new_txt and not re.search(r"\bactual\b", new_txt, flags=re.I):
                    last_guide = new_txt
            # Hide prior-year future annual initial rows when the first
            # in-year block already carries that prior value as Previous guide.
            for pos, rec in enumerate(ordered):
                annual_year = _annual_year_from_horizon(rec.get("Horizon"))
                stated_year, _ = _stated_quarter(rec.get("Stated in"))
                if annual_year is None or stated_year is None or stated_year >= annual_year:
                    continue
                later = [
                    cand for cand in ordered[pos + 1 :]
                    if _quarter_sort_score(cand.get("Display block")) >= annual_year * 10 + 1
                    and str(cand.get("Previous guide") or "").strip()
                ]
                if later:
                    rec["Hide from timeline"] = "1"
        for row in rows:
            actual_txt = str(row.get("Actual") or "").strip()
            metric_txt = str(row.get("Metric") or "").strip()
            progress_txt = str(row.get("Progress / run-rate") or "").strip()
            if actual_txt and _promise_value_looks_like_progress(actual_txt, metric=metric_txt):
                row["Progress / run-rate"] = progress_txt or _promise_progress_label(
                    actual_txt,
                    metric=metric_txt,
                    stated=row.get("Stated in"),
                )
                row["Actual"] = ""
                if str(row.get("Status") or "").strip().lower() in {"completed", "hit", "missed", "beat"}:
                    row["Status"] = "On track"
        return rows

    timeline_rows = _normalize_timeline_rows(timeline_rows)
    def _timeline_sort_key(rec: Mapping[str, Any]) -> Tuple[int, date]:
        stated_txt = str(rec.get("Stated in") or "")
        match = re.search(r"\b(20\d{2})-Q([1-4])\b", stated_txt)
        if match:
            period_score = int(match.group(1)) * 10 + int(match.group(2))
        else:
            horizon_txt = str(rec.get("Horizon") or "")
            match = re.search(r"\b(20\d{2})-Q([1-4])\b", horizon_txt)
            period_score = int(match.group(1)) * 10 + int(match.group(2)) if match else 0
        try:
            source_dt = pd.Timestamp(str(rec.get("Source date") or "")).date()
        except Exception:
            source_dt = date.min
        return (period_score, source_dt)

    timeline_rows.sort(key=_timeline_sort_key, reverse=True)

    latest_label = str(rows_sorted[0].get("block_label") or "")
    latest_rows = [r for r in rows_sorted if str(r.get("block_label") or "") == latest_label]
    open_rows: List[Dict[str, Any]] = []
    for r in latest_rows:
        horizon_txt = _horizon_from_row(r)
        status_txt = _canonical_promise_status(r.get("status"), metric_hint=r.get("metric"), horizon_hint=horizon_txt)
        actual_txt = str(r.get("latest") or "").strip()
        horizon_end = _horizon_end_from_label(horizon_txt)
        if status_txt not in {"Open", "On track"}:
            continue
        if horizon_end is not None and _promise_actual_for_horizon(r, horizon_txt, status_txt):
            continue
        open_rows.append(r)
    current_rows = latest_rows[:8]

    def _annual_guidance_year(label: Any) -> Optional[int]:
        m = re.fullmatch(r"(20\d{2})\s+year", str(label or "").strip(), flags=re.I)
        return int(m.group(1)) if m else None

    def _stated_quarter(label: Any) -> Tuple[Optional[int], Optional[int]]:
        m = re.search(r"\b(20\d{2})-Q([1-4])\b", str(label or ""), flags=re.I)
        if not m:
            return None, None
        return int(m.group(1)), int(m.group(2))

    open_annual_years = [_annual_guidance_year(_horizon_from_row(r)) for r in open_rows]
    open_annual_years = [year for year in open_annual_years if year is not None]
    current_open_year = max(open_annual_years) if open_annual_years else None

    annual_years = sorted(
        {
            year
            for year in (_annual_guidance_year(rec.get("Horizon")) for rec in timeline_rows)
            if year is not None and year != current_open_year
        },
        reverse=True,
    )

    def _annual_progression_rows(year: int) -> List[Dict[str, str]]:
        by_metric: Dict[str, Dict[str, str]] = {}
        metric_order: List[str] = []
        metric_latest_sort: Dict[str, Tuple[int, date]] = {}
        for rec in sorted(timeline_rows, key=_timeline_sort_key):
            if _annual_guidance_year(rec.get("Horizon")) != year:
                continue
            metric = str(rec.get("Metric") or "").strip()
            if not metric:
                continue
            if metric not in by_metric:
                by_metric[metric] = {
                    "Metric": metric,
                    "Initial guide": "",
                    "Q1 update": "",
                    "Q2 update": "",
                    "Q3 update": "",
                    "Q4 update": "",
                    "Actual": "",
                    "Status": "",
                    "Notes/source": "",
                }
                metric_order.append(metric)
            row = by_metric[metric]
            stated_year, stated_q = _stated_quarter(rec.get("Stated in"))
            if stated_year is not None and stated_q is not None:
                col_name = "Initial guide" if stated_year < year else f"Q{stated_q} update"
            else:
                col_name = "Initial guide" if not row.get("Initial guide") else "Q4 update"
            if col_name in row:
                row[col_name] = str(rec.get("New/current guide") or "")
            actual_txt = str(rec.get("Actual") or "").strip()
            if actual_txt and _has_measurable_actual_text(actual_txt):
                row["Actual"] = actual_txt
            status_txt = str(rec.get("Status") or "").strip()
            if status_txt:
                row["Status"] = status_txt
            note_txt = str(rec.get("Source / note") or "").strip()
            if note_txt:
                row["Notes/source"] = note_txt
            metric_latest_sort[metric] = _timeline_sort_key(rec)
        for metric, row in by_metric.items():
            final_actual = _annual_actual_for_metric(year, metric)
            if not final_actual:
                continue
            latest_guide = ""
            for col_name in ("Q4 update", "Q3 update", "Q2 update", "Q1 update", "Initial guide"):
                candidate = str(row.get(col_name) or "").strip()
                if candidate and not re.search(r"\bactual\b", candidate, flags=re.I):
                    latest_guide = candidate
                    break
            row["Actual"] = final_actual
            final_status = _status_from_guidance_actual(metric, latest_guide, final_actual)
            if final_status:
                row["Status"] = final_status
        metric_order.sort(key=lambda metric: metric_latest_sort.get(metric, (0, date.min)), reverse=True)
        return [by_metric[metric] for metric in metric_order]

    def _milestone_progression_rows(year: int) -> List[Dict[str, str]]:
        if ticker_txt != "GPRE":
            return []
        milestone_patterns = (
            "45z monetization",
            "advantage nebraska",
            "debt reduction",
            "cost savings target",
            "cost savings",
        )
        out: List[Dict[str, str]] = []
        seen: Set[Tuple[str, str]] = set()
        for rec in timeline_rows:
            metric_txt = str(rec.get("Metric") or "").strip()
            metric_low = metric_txt.lower()
            if not metric_txt or not any(pattern in metric_low for pattern in milestone_patterns):
                continue
            if _annual_guidance_year(rec.get("Horizon")) is not None:
                continue
            stated_year, _ = _stated_quarter(rec.get("Stated in"))
            horizon_year, _ = _stated_quarter(rec.get("Horizon"))
            if stated_year != year and horizon_year != year:
                continue
            status_txt = str(rec.get("Status") or "").strip()
            if status_txt.lower() in {"", "open", "not assessed"}:
                continue
            key = (metric_low, str(rec.get("Horizon") or "").strip().lower())
            if key in seen:
                continue
            seen.add(key)
            out.append(
                {
                    "Milestone": metric_txt,
                    "Target / plan": str(rec.get("New/current guide") or ""),
                    "Actual": str(rec.get("Actual") or rec.get("Progress / run-rate") or ""),
                    "Status": status_txt,
                    "Notes/source": str(rec.get("Source / note") or ""),
                }
            )
        return out

    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    blue = PatternFill("solid", fgColor="5B9BD5")
    title_blue = PatternFill("solid", fgColor="6FA8DC")
    header_fill = PatternFill("solid", fgColor="EAF3FB")
    neutral = PatternFill("solid", fgColor="FFFFFF")
    neutral_alt = PatternFill("solid", fgColor="F6F9FC")
    border = Border(bottom=Side(style="thin", color="D9E2EF"))

    def _status_fill(status: Any) -> PatternFill:
        low = _canonical_promise_status(status).lower()
        if low in {"completed"}:
            return PatternFill("solid", fgColor="009E73")
        if low in {"beat", "hit"}:
            return PatternFill("solid", fgColor="66C2A5")
        if low in {"on track"}:
            return PatternFill("solid", fgColor="56B4E9")
        if low in {"open"}:
            return PatternFill("solid", fgColor="A6CEE3")
        if low in {"mixed"}:
            return PatternFill("solid", fgColor="E69F00")
        if low in {"basis-dependent"}:
            return PatternFill("solid", fgColor="CC79A7")
        if low in {"missed"}:
            return PatternFill("solid", fgColor="D55E00")
        if low in {"not assessed"}:
            return PatternFill("solid", fgColor="D9D9D9")
        return PatternFill("solid", fgColor="E7EDF3")

    max_col = PROMISE_VISIBLE_MAX_COL
    row_idx = 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
    ws.cell(row_idx, 1, "Promise Progress").fill = title_blue
    ws.cell(row_idx, 1).font = Font(bold=True, size=15, color="FFFFFF")
    ws.cell(row_idx, 1).alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(1, max_col + 1):
        ws.cell(row_idx, cc).fill = title_blue
    ws.row_dimensions[row_idx].height = 26
    row_idx += 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
    ws.cell(row_idx, 1, f"{ticker_txt or 'Company'} guidance dashboard | newest periods first")
    ws.cell(row_idx, 1).font = Font(italic=True, size=10, color="666666")
    row_idx += 1

    def _section(title: str) -> None:
        nonlocal row_idx
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
        cell = ws.cell(row_idx, 1, title)
        cell.fill = blue
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for cc in range(1, max_col + 1):
            ws.cell(row_idx, cc).fill = blue
            ws.cell(row_idx, cc).border = border
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

    def _headers(labels: Sequence[str]) -> None:
        nonlocal row_idx
        for cc in range(1, max_col + 1):
            cell = ws.cell(row_idx, cc, labels[cc - 1] if cc <= len(labels) else "")
            cell.fill = header_fill
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

    def _write_row(values: Sequence[Any], *, status_col: Optional[int] = None, wrap_cols: Set[int] = frozenset()) -> None:
        nonlocal row_idx
        fill = neutral_alt if row_idx % 2 else neutral
        metric_hint = str(values[0] if values else "")
        for cc in range(1, max_col + 1):
            raw_val = values[cc - 1] if cc <= len(values) else ""
            val = _canonical_promise_status(raw_val, metric_hint=metric_hint) if status_col == cc else _format_promise_value_text(raw_val, metric_hint)
            cell = ws.cell(row_idx, cc, val)
            cell.fill = _status_fill(val) if status_col == cc else fill
            cell.font = Font(size=11)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc in wrap_cols)
        ws.row_dimensions[row_idx].height = 24
        row_idx += 1

    _section("Management Credibility Scorecard")
    _headers(["Category", "Score", "Evidence", "", "", "", "Read"])
    ws.merge_cells(start_row=row_idx - 1, start_column=3, end_row=row_idx - 1, end_column=6)
    ws.merge_cells(start_row=row_idx - 1, start_column=7, end_row=row_idx - 1, end_column=max_col)
    for category, score, evidence, read in _management_credibility_scorecard_rows(ticker_txt):
        data_row = row_idx
        _write_row([category, score, evidence, "", "", "", read], wrap_cols={3, 7})
        ws.merge_cells(start_row=data_row, start_column=3, end_row=data_row, end_column=6)
        ws.merge_cells(start_row=data_row, start_column=7, end_row=data_row, end_column=max_col)
        ws.cell(data_row, 3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(data_row, 7).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    row_idx += 1

    for annual_year in annual_years:
        annual_rows = _annual_progression_rows(annual_year)
        if not annual_rows:
            continue
        _section(f"{annual_year} guidance progression")
        header_row = row_idx
        _headers(["Metric", "Initial guide", "Q1 update", "Q2 update", "Q3 update", "Q4 update", "Actual", "Status", "Notes/source"])
        ws.merge_cells(start_row=header_row, start_column=9, end_row=header_row, end_column=max_col)
        ws.cell(header_row, 9).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        for annual_row in annual_rows:
            data_row = row_idx
            _write_row(
                [
                    annual_row.get("Metric"),
                    annual_row.get("Initial guide"),
                    annual_row.get("Q1 update"),
                    annual_row.get("Q2 update"),
                    annual_row.get("Q3 update"),
                    annual_row.get("Q4 update"),
                    annual_row.get("Actual"),
                    annual_row.get("Status"),
                    annual_row.get("Notes/source"),
                ],
                status_col=8,
                wrap_cols={9},
            )
            ws.merge_cells(start_row=data_row, start_column=9, end_row=data_row, end_column=max_col)
            ws.cell(data_row, 9).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row_idx += 1

    for milestone_year in [2025]:
        milestone_rows = _milestone_progression_rows(milestone_year)
        if not milestone_rows:
            continue
        _section(f"{milestone_year} milestone progression")
        header_row = row_idx
        _headers(["Milestone", "Target / plan", "Actual", "Status", "Notes/source"])
        ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=max_col)
        ws.cell(header_row, 5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        for milestone_row in milestone_rows:
            data_row = row_idx
            _write_row(
                [
                    milestone_row.get("Milestone"),
                    milestone_row.get("Target / plan"),
                    milestone_row.get("Actual"),
                    milestone_row.get("Status"),
                    milestone_row.get("Notes/source"),
                ],
                status_col=4,
                wrap_cols={5},
            )
            ws.merge_cells(start_row=data_row, start_column=5, end_row=data_row, end_column=max_col)
            ws.cell(data_row, 5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row_idx += 1

    if ticker_txt == "PBI" and current_open_year == 2026:
        pbi_open_guides = {
            "Revenue guidance": ("$1.8bn-$1.86bn", "2026 year Revenue guidance updated to $1.8bn-$1.86bn."),
            "Adjusted EBIT guidance": ("$425m-$465m", "2026 year Adjusted EBIT guidance $425m-$465m."),
            "Adjusted EPS guidance": ("$1.50-$1.65", "2026 year Adjusted EPS guidance $1.50-$1.65."),
            "FCF target": ("$345m-$380m", "2026 year source-defined Free Cash Flow target $345m-$380m."),
        }
        for row in open_rows:
            metric_txt = str(row.get("metric") or "").strip()
            if metric_txt not in pbi_open_guides:
                continue
            target_txt, note_txt = pbi_open_guides[metric_txt]
            row["target"] = target_txt
            row["note"] = note_txt
            row["status"] = "Open"

    open_section_title = f"{current_open_year} open guidance" if current_open_year else "Open guidance"
    _section(open_section_title)
    header_row = row_idx
    _headers(["Metric", "Current guide", "Horizon", "Status", "Notes/source"])
    ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=max_col)
    ws.cell(header_row, 5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for row in open_rows:
        data_row = row_idx
        _write_row(
            [row.get("metric"), row.get("target"), _horizon_from_row(row), row.get("status"), row.get("note")],
            status_col=4,
            wrap_cols={5},
        )
        ws.merge_cells(start_row=data_row, start_column=5, end_row=data_row, end_column=max_col)
        ws.cell(data_row, 5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row_idx += 1
    _section("Quarterly guidance timeline / revision log")
    timeline_headers = PROMISE_TIMELINE_HEADERS
    def _is_prior_period_current_open_annual_row(rec: Mapping[str, Any]) -> bool:
        if current_open_year is None:
            return False
        if _annual_guidance_year(rec.get("Horizon")) != current_open_year:
            return False
        stated_year, _stated_q = _stated_quarter(rec.get("Stated in"))
        if stated_year is None or stated_year >= current_open_year:
            return False
        return str(rec.get("Change type") or "").strip().lower() == "maintained"

    display_timeline_rows = [
        rec for rec in timeline_rows
        if not _is_prior_period_current_open_annual_row(rec)
        and str(rec.get("Hide from timeline") or "").strip() != "1"
    ]

    def _quarter_score(label: Any) -> int:
        m = re.search(r"\b(20\d{2})-Q([1-4])\b", str(label or ""), flags=re.I)
        if m:
            return int(m.group(1)) * 10 + int(m.group(2))
        return 0

    def _source_ordinal(rec: Mapping[str, Any]) -> int:
        try:
            return pd.Timestamp(str(rec.get("Source date") or "")).date().toordinal()
        except Exception:
            return 0

    def _horizon_group_sort_key(rec: Mapping[str, Any]) -> Tuple[int, int, str]:
        stated_score = _quarter_score(rec.get("Display block") or rec.get("Stated in"))
        horizon_txt = str(rec.get("Horizon") or "").strip()
        annual_year = _annual_guidance_year(horizon_txt)
        if annual_year is not None:
            return (0, -(annual_year * 10 + 5), str(rec.get("Metric") or ""))
        horizon_score = _quarter_score(horizon_txt)
        if horizon_score and stated_score and horizon_score > stated_score:
            rank = 1
        elif horizon_score and stated_score and horizon_score == stated_score:
            rank = 2
        elif horizon_score:
            rank = 3
        else:
            rank = 4
        return (rank, -horizon_score, str(rec.get("Metric") or ""))

    def _display_timeline_sort_key(rec: Mapping[str, Any]) -> Tuple[int, int, int, int, str]:
        stated_score = _quarter_score(rec.get("Display block") or rec.get("Stated in"))
        horizon_rank, horizon_score, metric = _horizon_group_sort_key(rec)
        return (-stated_score, -_source_ordinal(rec), horizon_rank, horizon_score, metric)

    display_timeline_rows = sorted(display_timeline_rows, key=_display_timeline_sort_key)
    last_group = ""
    for rec in display_timeline_rows:
        group = str(rec.get("Display block") or rec.get("Stated in") or "")
        if group != last_group:
            if last_group:
                row_idx += 1
            _section(f"{group} revisions")
            _headers(timeline_headers)
            last_group = group
        data_row = row_idx
        _write_row([rec.get(h, "") for h in timeline_headers], status_col=7, wrap_cols={11})
        try:
            ws.merge_cells(start_row=data_row, start_column=11, end_row=data_row, end_column=max_col)
            ws.cell(data_row, 11).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        except ValueError:
            pass
        hidden_id = _promise_lifecycle_id(rec.get("Metric"), rec.get("Horizon"))
        if hidden_id:
            ws.cell(row_idx - 1, 15).value = hidden_id
            ws.column_dimensions["O"].hidden = True

    for merge_range in list(ws.merged_cells.ranges):
        if merge_range.min_row == merge_range.max_row:
            row_first = str(ws.cell(merge_range.min_row, 1).value or "").strip()
            if row_first in {"Metric", "Milestone", "Category"} or str(ws.cell(merge_range.min_row, 1).fill.fgColor.rgb or "").upper().endswith(("5B9BD5", "6FA8DC", "4472C4")):
                ws.unmerge_cells(str(merge_range))
    current_section_for_merge = ""
    for merge_row in range(1, int(ws.max_row or 0) + 1):
        first_value = str(ws.cell(merge_row, 1).value or "").strip()
        first_fill = str(ws.cell(merge_row, 1).fill.fgColor.rgb or "").upper()
        if first_value and first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")):
            current_section_for_merge = first_value
            ws.merge_cells(start_row=merge_row, start_column=1, end_row=merge_row, end_column=max_col)
        elif first_value == "Category":
            labels = ["Category", "Score", "Evidence", "", "", "", "Read", "", "", "", "", ""]
            for col_idx, label in enumerate(labels, start=1):
                ws.cell(merge_row, col_idx).value = label
            ws.merge_cells(start_row=merge_row, start_column=3, end_row=merge_row, end_column=6)
            ws.merge_cells(start_row=merge_row, start_column=7, end_row=merge_row, end_column=max_col)
        elif first_value == "Metric":
            if current_section_for_merge.endswith("guidance progression"):
                labels = ["Metric", "Initial guide", "Q1 update", "Q2 update", "Q3 update", "Q4 update", "Actual", "Status", "Notes/source", "", "", ""]
            elif current_section_for_merge.endswith("open guidance") or current_section_for_merge == "Open guidance":
                labels = ["Metric", "Current guide", "Horizon", "Status", "Notes/source", "", "", "", "", "", "", ""]
            else:
                labels = list(PROMISE_TIMELINE_HEADERS) + [""]
            for col_idx, label in enumerate(labels, start=1):
                ws.cell(merge_row, col_idx).value = label
            row_values = [str(ws.cell(merge_row, col_idx).value or "").strip() for col_idx in range(1, max_col + 1)]
            if "Notes/source" in row_values:
                note_col_idx = row_values.index("Notes/source") + 1
                if note_col_idx < max_col:
                    ws.merge_cells(start_row=merge_row, start_column=note_col_idx, end_row=merge_row, end_column=max_col)
        elif first_value == "Milestone":
            labels = ["Milestone", "Target / plan", "Actual", "Status", "Notes/source", "", "", "", "", "", "", ""]
            for col_idx, label in enumerate(labels, start=1):
                ws.cell(merge_row, col_idx).value = label
            ws.merge_cells(start_row=merge_row, start_column=5, end_row=merge_row, end_column=max_col)

    widths = {
        "A": 28,
        "B": 28,
        "C": 32,
        "D": 15,
        "E": 22,
        "F": 28,
        "G": 15,
        "H": 14,
        "I": 16,
        "J": 14,
        "K": 42,
        "L": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    _remove_empty_promise_revision_blocks(ws)
    _polish_promise_scorecard_layout(ws)
    parent_wb = getattr(ws, "parent", None)
    _finalize_promise_revision_semantics(ws)
    if parent_wb is not None:
        _apply_source_backed_promise_mapping_overrides(parent_wb, ticker)
    _apply_promise_grid_style(ws)
    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 112


