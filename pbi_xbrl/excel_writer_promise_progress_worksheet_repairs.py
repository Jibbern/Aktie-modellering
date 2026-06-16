"""Post-render Promise_Progress_UI worksheet repair and polish helpers."""
from __future__ import annotations

import math
import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from typing import Any, Dict, List, Mapping, MutableMapping, Optional, Set, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.workbook import Workbook


@dataclass(frozen=True)
class PromiseProgressWorksheetRepairDeps:
    runtime: MutableMapping[str, Any]


_runtime: MutableMapping[str, Any] = {}


def _set_runtime(deps: PromiseProgressWorksheetRepairDeps) -> None:
    global _runtime
    _runtime = deps.runtime


def _runtime_call(name: str, *args: Any, **kwargs: Any) -> Any:
    return _runtime[name](*args, **kwargs)


def _date_or_none(*args: Any, **kwargs: Any) -> Any:
    return _runtime_call("_date_or_none", *args, **kwargs)


def _date_is_missing_or_outside(*args: Any, **kwargs: Any) -> Any:
    return _runtime_call("_date_is_missing_or_outside", *args, **kwargs)


def _shared_visible_period_text(*args: Any, **kwargs: Any) -> str:
    return _runtime_call("_shared_visible_period_text", *args, **kwargs)


def _apply_source_backed_promise_mapping_overrides(*args: Any, **kwargs: Any) -> Any:
    return _runtime_call("_apply_source_backed_promise_mapping_overrides", *args, **kwargs)


def _management_credibility_scorecard_rows(ticker: Any = "") -> List[Tuple[str, str, str, str]]:
    ticker_txt = str(ticker or "").strip().upper()
    if ticker_txt == "ANF":
        return [
            ("Sales guidance accuracy", "Good", "2025 sales guide was evaluated against reported results; 2026 remains open.", "Demand guide credibility is useful, but comps are tougher."),
            ("Margin guidance accuracy", "Mixed", "2025 GAAP vs adjusted basis matters; 2026 margin bridge has tariff/ERP/freight moving pieces.", "Margin proof is the main debate."),
            ("EPS guidance accuracy", "Basis-dependent", "Adjusted EPS and GAAP EPS differ, so basis discipline is required.", "Use adjusted guide for promise tracking, with GAAP context visible."),
            ("Buyback/capital allocation delivery", "Strong", "2025 buybacks were about $450m / 5.4m shares.", "Capital returns supported EPS while net cash stayed strong."),
            ("Inventory discipline", "Good", "Inventory growth has tariff and ERP prebuild explanations.", "Still watch markdown risk if demand slows."),
        ]
    if ticker_txt == "PBI":
        return [
            ("FCF delivery", "Mixed", "FCF is central to the turnaround and debt case.", "Cash conversion must remain durable, not one-quarter noise."),
            ("Cost savings delivery", "Good", "Current savings target/run-rate is tracked in Promise and Valuation.", "Execution is credible if savings continue to flow through EBIT."),
            ("Debt/refinancing execution", "Watch", "Debt and maturities remain core diligence items.", "Equity case needs lower refinancing risk."),
            ("Segment stabilization", "Mixed", "Presort and SendTech trends drive recurring earnings quality.", "Segment proof remains necessary."),
            ("Guidance transparency", "Good", "PBI uses curated Slides_Guidance / guidance profile for clean user-facing guidance.", "Clean-empty normalized guidance table is intentional until a reliable normalizer exists."),
        ]
    if ticker_txt == "GPRE":
        return [
            ("45Z delivery", "On track", "45Z contribution and facility status are tracked in Promise and Valuation.", "Policy upside is meaningful but still needs cash conversion."),
            ("Capex discipline", "Good", "Capex guide is visible and tracked as open/current guidance.", "FCF depends on capex staying disciplined through the cycle."),
            ("Margin/commodity commentary accuracy", "Watch", "Crush margin and input spreads are cyclical and source-sensitive.", "Management credibility depends on margin commentary matching realized economics."),
            ("Policy execution", "Mixed", "RFS/RVO/E15/export and 45Z milestones can move the case.", "Policy rows are qualitative where source-backed, not over-scored."),
            ("FCF/balance sheet discipline", "Watch", "Debt/cash/liquidity are visible in Valuation and Debt Detail.", "Commodity upside needs to convert into balance-sheet resilience."),
        ]
    return [
        ("Guidance delivery", "Not enough data", "Promise history is sparse.", "Use this section as a compact credibility checklist."),
    ]

def _insert_management_credibility_scorecard(ws: Any, ticker: Any = "") -> None:
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    for row in ws.iter_rows(min_row=1, max_row=min(int(ws.max_row or 0), 30), min_col=1, max_col=min(int(ws.max_column or 0), 10)):
        if any(str(cell.value or "").strip() == "Management Credibility Scorecard" for cell in row):
            return
    rows = _management_credibility_scorecard_rows(ticker)
    if not rows:
        return
    start_row = 3
    insert_count = len(rows) + 3
    ws.insert_rows(start_row, insert_count)
    blue = PatternFill("solid", fgColor="5B9BD5")
    header_fill = PatternFill("solid", fgColor="EAF3FB")
    neutral_fill = PatternFill("solid", fgColor="FFFFFF")
    neutral_alt = PatternFill("solid", fgColor="F6F9FC")
    thin = Border(left=Side(style="thin", color="D9E2EA"), right=Side(style="thin", color="D9E2EA"), top=Side(style="thin", color="D9E2EA"), bottom=Side(style="thin", color="D9E2EA"))
    dark = "1F2933"

    def _fill_row(row_idx: int, fill: PatternFill) -> None:
        for cc in range(1, 11):
            ws.cell(row_idx, cc).fill = copy(fill)
            ws.cell(row_idx, cc).border = thin

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
    _fill_row(start_row, blue)
    ws.cell(start_row, 1, "Management Credibility Scorecard")
    ws.cell(start_row, 1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(start_row, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row].height = 22

    header_row = start_row + 1
    _fill_row(header_row, header_fill)
    for cc, label in enumerate(("Category", "Score", "Evidence", "Read"), start=1):
        target_col = {1: 1, 2: 2, 3: 3, 4: 7}[cc]
        ws.cell(header_row, target_col, label)
        ws.cell(header_row, target_col).font = Font(bold=True, size=11, color=dark)
        ws.cell(header_row, target_col).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=header_row, start_column=3, end_row=header_row, end_column=6)
    ws.merge_cells(start_row=header_row, start_column=7, end_row=header_row, end_column=10)
    ws.row_dimensions[header_row].height = 22

    row_idx = header_row + 1
    for idx, (category, score, evidence, read) in enumerate(rows):
        fill = neutral_alt if idx % 2 == 0 else neutral_fill
        _fill_row(row_idx, fill)
        vals = {1: category, 2: score, 3: evidence, 7: read}
        for cc, value in vals.items():
            cell = ws.cell(row_idx, cc, _shared_visible_period_text(str(value or "")))
            cell.font = Font(size=11, color=dark)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc in {3, 7})
        ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=6)
        ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=10)
        ws.row_dimensions[row_idx].height = 24
        row_idx += 1
    ws.column_dimensions["A"].width = max(float(ws.column_dimensions["A"].width or 0), 36.0)
    ws.column_dimensions["B"].width = max(float(ws.column_dimensions["B"].width or 0), 14.0)
    ws.freeze_panes = "A2"

def _polish_promise_scorecard_layout(ws: Any) -> None:
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    _remove_blank_promise_rows(ws)
    _remove_actual_only_promise_rows(ws)
    _ensure_cost_savings_run_rate_revision_row(ws)
    _dedupe_promise_progress_rows(ws)
    _remove_promise_metric_stubs(ws)
    _ensure_q4_annual_actual_revision_rows(ws)
    _cleanup_anf_promise_after_repair(ws)
    _clear_pre_release_promise_actuals(ws)
    _standardize_promise_section_layout(ws)
    _repair_promise_table_header_merges(ws)
    _remove_promise_metric_stubs(ws)
    _remove_pbi_duplicate_cost_savings_timeline_rows(ws)
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)

    def _repair_pbi_open_guidance_values() -> None:
        header_blob = " ".join(
            str(ws.cell(rr, cc).value or "")
            for rr in range(1, min(int(ws.max_row or 0), 4) + 1)
            for cc in range(1, max_col + 1)
        )
        if "PBI guidance dashboard" not in header_blob:
            return
        current_section = ""
        active_cols: Dict[str, int] = {}
        pbi_guides = {
            "Revenue guidance": ("$1.8bn-$1.86bn", "2026 year Revenue guidance updated to $1.8bn-$1.86bn."),
            "Adjusted EBIT guidance": ("$425m-$465m", "2026 year Adjusted EBIT guidance $425m-$465m."),
            "Adjusted EPS guidance": ("$1.50-$1.65", "2026 year Adjusted EPS guidance $1.50-$1.65."),
            "FCF target": ("$345m-$380m", "2026 year source-defined Free Cash Flow target $345m-$380m."),
        }
        for rr in range(1, int(ws.max_row or 0) + 1):
            first_txt = str(ws.cell(rr, 1).value or "").strip()
            if _is_promise_section_row(ws, rr):
                current_section = first_txt
                active_cols = {}
                continue
            row_map = {
                str(ws.cell(rr, cc).value or "").strip().lower(): cc
                for cc in range(1, max_col + 1)
                if str(ws.cell(rr, cc).value or "").strip()
            }
            if "metric" in row_map and ("current guide" in row_map or "new/current guide" in row_map):
                active_cols = row_map
                continue
            if "open guidance" not in current_section.lower() or not active_cols:
                continue
            metric_col = active_cols.get("metric")
            if not metric_col:
                continue
            metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
            if metric_txt not in pbi_guides:
                continue
            guide, note = pbi_guides[metric_txt]
            guide_col = active_cols.get("current guide") or active_cols.get("new/current guide")
            horizon_col = active_cols.get("horizon")
            status_col = active_cols.get("status")
            note_col = active_cols.get("notes/source") or active_cols.get("source / note")
            if guide_col:
                ws.cell(rr, guide_col).value = guide
            if horizon_col:
                ws.cell(rr, horizon_col).value = "2026 year"
            if status_col:
                ws.cell(rr, status_col).value = "Open"
            if note_col:
                ws.cell(rr, note_col).value = note

    if max_col < 10:
        return
    scorecard_row: Optional[int] = None
    for rr in range(1, min(int(ws.max_row or 0), 20) + 1):
        if str(ws.cell(rr, 1).value or "").strip() == "Management Credibility Scorecard":
            scorecard_row = rr
            break
    if scorecard_row is None:
        _finalize_promise_revision_semantics(ws)
        _repair_promise_table_header_merges(ws)
        _repair_pbi_open_guidance_values()
        _repair_anf_promise_actual_progress_semantics(ws)
        _apply_promise_grid_style(ws)
        return

    def _has_merge(row_idx: int, start_col: int, end_col: int) -> bool:
        return any(
            merge_range.min_row == row_idx
            and merge_range.max_row == row_idx
            and merge_range.min_col == start_col
            and merge_range.max_col == end_col
            for merge_range in ws.merged_cells.ranges
        )

    header_row = scorecard_row + 1
    if str(ws.cell(header_row, 1).value or "").strip() == "Category":
        if not _has_merge(header_row, 3, 6):
            ws.merge_cells(start_row=header_row, start_column=3, end_row=header_row, end_column=6)
        if not _has_merge(header_row, 7, 10):
            ws.merge_cells(start_row=header_row, start_column=7, end_row=header_row, end_column=10)
        ws.row_dimensions[header_row].height = 22

    ws.row_dimensions[scorecard_row].height = 22
    row_idx = header_row + 1
    while row_idx <= int(ws.max_row or 0):
        first_txt = str(ws.cell(row_idx, 1).value or "").strip()
        first_fill = str(ws.cell(row_idx, 1).fill.fgColor.rgb or "").upper()
        if not first_txt:
            break
        if first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")) or first_txt.endswith("progression") or first_txt.endswith("guidance"):
            break
        if not _has_merge(row_idx, 3, 6):
            ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=6)
        if not _has_merge(row_idx, 7, 10):
            ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=10)
        for col_idx in (1, 2, 3, 7):
            ws.cell(row_idx, col_idx).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=col_idx in {3, 7},
            )
        ws.row_dimensions[row_idx].height = 24
        row_idx += 1
    _finalize_promise_revision_semantics(ws)
    _repair_promise_table_header_merges(ws)
    _repair_pbi_open_guidance_values()
    _remove_pbi_duplicate_cost_savings_timeline_rows(ws)
    _repair_anf_promise_actual_progress_semantics(ws)
    _apply_promise_grid_style(ws)

def _promise_header_name(value: Any) -> str:
    txt = str(value or "").strip().lower()
    if txt in {"actual / latest actual", "actual / latest", "latest actual", "latest result"}:
        return "actual"
    return txt

def _set_promise_row_semantics(
    ws: Any,
    row_idx: int,
    cols: Mapping[str, int],
    *,
    change_type: Any = None,
    actual: Any = None,
    progress: Any = None,
    status: Any = None,
    note: Any = None,
) -> None:
    """Apply the shared visible Promise Actual/Progress convention to one row."""
    updates = (
        ("change type", change_type),
        ("actual", actual),
        ("progress / run-rate", progress),
        ("status", status),
    )
    for header_key, value in updates:
        if value is None:
            continue
        col_idx = cols.get(header_key)
        if col_idx:
            ws.cell(row_idx, col_idx).value = value
    if note is not None:
        note_col = cols.get("source / note") or cols.get("notes/source")
        if note_col:
            ws.cell(row_idx, note_col).value = note

PROMISE_TIMELINE_HEADERS = [
    "Metric",
    "Previous guide",
    "New/current guide",
    "Change type",
    "Actual",
    "Progress / run-rate",
    "Status",
    "Horizon",
    "Stated in",
    "Source date",
    "Source / note",
]

PROMISE_VISIBLE_MAX_COL = 12

def _promise_hidden_key_slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")

def _ensure_anf_promise_hidden_source_keys(ws: Any) -> None:
    """Keep ANF timeline trace keys aligned after last-mile visible repairs.

    ANF source-backed rows are curated in several post-processing passes so that
    Q4 Actual and FY/YTD Progress remain visibly correct. Some of those passes
    write cells directly instead of going through the generic upsert helper that
    normally fills hidden source keys. This pass only fills column O with a
    stable derived guidance key for source-backed ANF timeline rows.
    """
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    wb_obj = getattr(ws, "parent", None)
    if wb_obj is None or not any(str(name).startswith("ANF_") for name in getattr(wb_obj, "sheetnames", [])):
        return

    max_row = int(ws.max_row or 0)
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)
    current_section = ""
    active_cols: Dict[str, int] = {}
    for rr in range(1, max_row + 1):
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        if _is_promise_section_row(ws, rr):
            current_section = first_txt
            active_cols = {}
            continue
        row_map = {
            _promise_header_name(ws.cell(rr, cc).value): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(rr, cc).value or "").strip()
        }
        if "metric" in row_map and ("source date" in row_map or "source / note" in row_map):
            active_cols = row_map
            continue
        if not active_cols or not current_section.endswith("revisions"):
            continue
        metric_col = active_cols.get("metric") or active_cols.get("milestone")
        if not metric_col:
            continue
        metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
        if not metric_txt or metric_txt.lower() in {"metric", "milestone"}:
            continue
        source_date_col = active_cols.get("source date")
        note_col = active_cols.get("source / note") or active_cols.get("notes/source")
        source_date_txt = str(ws.cell(rr, source_date_col).value or "").strip() if source_date_col else ""
        note_txt = str(ws.cell(rr, note_col).value or "").strip() if note_col else ""
        if not source_date_txt and not note_txt:
            continue
        horizon_col = active_cols.get("horizon")
        stated_col = active_cols.get("stated in")
        key_parts = [
            "guidance",
            "anf",
            _promise_hidden_key_slug(metric_txt),
            _promise_hidden_key_slug(ws.cell(rr, horizon_col).value if horizon_col else ""),
            _promise_hidden_key_slug(ws.cell(rr, stated_col).value if stated_col else ""),
            _promise_hidden_key_slug(source_date_txt),
        ]
        hidden_key = ":".join(part for part in key_parts if part)
        if hidden_key and hidden_key != "guidance:anf":
            existing_key = str(ws.cell(rr, 15).value or "").strip()
            if existing_key != hidden_key:
                ws.cell(rr, 15).value = hidden_key
    ws.column_dimensions["O"].hidden = True

def _promise_stated_quarter_parts(label: Any) -> Tuple[Optional[int], Optional[int]]:
    m = re.search(r"\b(20\d{2})-Q([1-4])\b", str(label or ""), flags=re.I)
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))

def _promise_annual_year(label: Any) -> Optional[int]:
    m = re.fullmatch(r"(20\d{2})\s+year", str(label or "").strip(), flags=re.I)
    return int(m.group(1)) if m else None

def _promise_progress_label(value: Any, *, metric: Any = "", stated: Any = "") -> str:
    txt = str(value or "").strip()
    if not txt:
        return ""
    if re.search(r"\b(run[- ]rate|implemented|identified|achieved|realized|monetized|ytd|ttm|q[1-4]:|progress|operational|qualified|of\s+\d+)\b", txt, flags=re.I):
        if re.search(r"^\$?\s*\d", txt) and re.search(r"\brun[- ]rate\b", txt, flags=re.I) and not txt.lower().startswith("run-rate:"):
            amount_match = re.search(r"\$?\s*(\d+(?:\.\d+)?)\s*m", txt, flags=re.I)
            return f"Run-rate: ${float(amount_match.group(1)):g}m" if amount_match else f"Run-rate: {txt}"
        return txt
    _, stated_q = _promise_stated_quarter_parts(stated)
    if stated_q in {1, 2, 3}:
        return f"Q{stated_q}: {txt}"
    return txt

def _promise_value_looks_like_progress(value: Any, *, metric: Any = "") -> bool:
    txt = str(value or "").strip()
    metric_low = str(metric or "").strip().lower()
    if not txt:
        return False
    if any(token in metric_low for token in ("startup", "strategic milestone")) and re.search(r"\boperational\b", txt, flags=re.I):
        return False
    if "facility qualification" in metric_low and re.search(r"\b(all\s+8|all\s+eight|8\s+of\s+8)\b", txt, flags=re.I) and re.search(
        r"\b(qualified|qualifying|operational|running)\b",
        txt,
        flags=re.I,
    ):
        return False
    if any(token in metric_low for token in ("cost savings", "facility qualification")):
        return True
    return bool(re.search(r"\b(run[- ]rate|implemented|identified|achieved|realized|monetized|ytd|ttm|progress|operational|qualified|of\s+\d+)\b", txt, flags=re.I))

def _promise_revision_event_from_section(section: Any) -> str:
    txt = str(section or "").strip()
    return re.sub(r"\s+revisions\s*$", "", txt, flags=re.I).strip()

def _promise_event_sort_key(value: Any, source_date: Any = "") -> Tuple[date, int, str]:
    source_txt = str(source_date or "").strip()
    try:
        source_dt = pd.Timestamp(source_txt).date() if source_txt else date.min
    except Exception:
        source_dt = date.min
    event_txt = str(value or "").strip()
    m = re.search(r"\b(20\d{2})-Q([1-4])\b", event_txt, flags=re.I)
    if m:
        score = int(m.group(1)) * 10 + int(m.group(2))
        if "pre-release" in event_txt.lower():
            score = score * 10 + 5
        else:
            score = score * 10 + 9
    else:
        score = 0
    return source_dt, score, event_txt.lower()

def _promise_metric_definition_key(value: Any) -> str:
    """Normalize Promise metrics enough for previous-guide matching.

    Previous guide should compare like with like: same economic metric and same
    accounting definition. Keep this conservative so adjusted and GAAP/free-cash
    variants do not borrow each other's prior guide.
    """
    txt = str(value or "").strip().lower()
    if not txt:
        return ""
    if "adjusted fcf" in txt or "adjusted free cash" in txt:
        return "adjusted_fcf"
    if "free cash" in txt or re.search(r"\bfcf\b", txt):
        return "fcf"
    if "adjusted" in txt and "eps" in txt:
        return "adjusted_eps"
    if "gaap" in txt and "eps" in txt:
        return "gaap_eps"
    if "diluted" in txt and "eps" in txt:
        return "diluted_eps"
    if "eps" in txt:
        return "eps"
    if "adjusted ebitda" in txt or "adj ebitda" in txt:
        return "adjusted_ebitda"
    if "ebitda" in txt:
        return "ebitda"
    if "adjusted ebit" in txt or "adj ebit" in txt:
        return "adjusted_ebit"
    if re.search(r"\bebit\b", txt):
        return "ebit"
    if "operating margin" in txt or "margin" in txt:
        return "operating_margin"
    if "revenue" in txt or "sales" in txt:
        return "revenue"
    if "capex" in txt or "capital expenditure" in txt:
        return "capex"
    if "buyback" in txt or "repurchase" in txt:
        return "buybacks"
    if "share count" in txt or "diluted shares" in txt:
        return "shares"
    if "cost savings" in txt or "cost reduction" in txt or "restructuring" in txt:
        return "cost_savings"
    if "debt" in txt or "liquidity" in txt or "leverage" in txt:
        return "debt_liquidity"
    if "45z" in txt:
        return f"45z::{txt}"
    return re.sub(r"\s+", " ", txt)

def _promise_metric_order_rank(value: Any) -> int:
    txt = str(value or "").strip().lower()
    if not txt:
        return 999
    if "revenue" in txt or "sales" in txt:
        return 10
    if "operating margin" in txt or ("margin" in txt and "bridge" not in txt):
        return 20
    if "ebitda" in txt:
        return 30
    if re.search(r"\bebit\b", txt):
        return 40
    if "eps" in txt:
        return 50
    if "fcf" in txt or "free cash" in txt or "cash flow" in txt:
        return 60
    if "capex" in txt or "capital expenditure" in txt:
        return 70
    if "buyback" in txt or "repurchase" in txt or "share count" in txt or "diluted shares" in txt:
        return 80
    if "cost savings" in txt or "cost reduction" in txt or "restructuring" in txt or "cash optimization" in txt:
        return 90
    if "debt" in txt or "leverage" in txt or "liquidity" in txt:
        return 100
    if "45z" in txt or "policy" in txt or "facility" in txt or "segment" in txt or "advantage nebraska" in txt:
        return 110
    return 900

def _clean_gpre_45z_monetization_value(value: Any) -> Any:
    txt = str(value or "").strip()
    if not txt:
        return value
    range_match = re.search(r"\$?\s*(\d+(?:\.\d+)?)\s*m\s*[-–]\s*\$?\s*(\d+(?:\.\d+)?)\s*m", txt, flags=re.I)
    if range_match:
        return f"${range_match.group(1)}m-${range_match.group(2)}m"
    single_match = re.search(r"\$?\s*(\d+(?:\.\d+)?)\s*m\b", txt, flags=re.I)
    if single_match and len(txt) > len(single_match.group(0)) + 6:
        return f"${single_match.group(1)}m"
    return value

def _finalize_promise_revision_semantics(ws: Any) -> None:
    """Last semantic guard for Promise timeline rows.

    Revision blocks are event/stated-in blocks.  This pass removes rows that
    drifted into the wrong block, removes actual-only rows, fills missing prior
    guides from earlier same metric+horizon events, and keeps GPRE 45Z visible
    values concise without losing source notes.
    """
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)

    def _header_map(row_idx: int) -> Dict[str, int]:
        return {
            _promise_header_name(ws.cell(row_idx, cc).value): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(row_idx, cc).value or "").strip()
        }

    def _timeline_rows() -> List[Dict[str, Any]]:
        out: List[Dict[str, Any]] = []
        active_cols: Dict[str, int] = {}
        section = ""
        for rr in range(1, int(ws.max_row or 0) + 1):
            first_txt = str(ws.cell(rr, 1).value or "").strip()
            if _is_promise_section_row(ws, rr):
                section = first_txt
                active_cols = {}
                continue
            row_map = _header_map(rr)
            if {"metric", "previous guide", "new/current guide", "change type", "actual", "status"}.issubset(set(row_map)):
                active_cols = row_map
                continue
            if not active_cols or not section.endswith("revisions"):
                continue
            metric_col = active_cols.get("metric")
            if not metric_col:
                continue
            metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
            if not metric_txt or metric_txt.lower() == "metric":
                continue
            out.append({"row": rr, "section": section, "cols": dict(active_cols)})
        return out

    rows_to_delete: Set[int] = set()
    for item in _timeline_rows():
        rr = int(item["row"])
        section = str(item["section"])
        cols = item["cols"]
        metric_col = cols.get("metric")
        prev_col = cols.get("previous guide")
        new_col = cols.get("new/current guide")
        change_col = cols.get("change type")
        actual_col = cols.get("actual")
        progress_col = cols.get("progress / run-rate")
        stated_col = cols.get("stated in")
        note_col = cols.get("source / note")
        metric_txt = str(ws.cell(rr, metric_col).value or "").strip() if metric_col else ""
        prev_txt = str(ws.cell(rr, prev_col).value or "").strip() if prev_col else ""
        new_txt = str(ws.cell(rr, new_col).value or "").strip() if new_col else ""
        change_txt = str(ws.cell(rr, change_col).value or "").strip().lower() if change_col else ""
        actual_txt = str(ws.cell(rr, actual_col).value or "").strip() if actual_col else ""
        progress_txt = str(ws.cell(rr, progress_col).value or "").strip() if progress_col else ""
        stated_txt = str(ws.cell(rr, stated_col).value or "").strip() if stated_col else ""
        note_txt = str(ws.cell(rr, note_col).value or "").strip() if note_col else ""
        event_txt = _promise_revision_event_from_section(section)
        stated_low = stated_txt.strip().lower()
        event_low = event_txt.strip().lower()
        if stated_low and event_low and stated_low != event_low and stated_low not in event_low:
            rows_to_delete.add(rr)
            continue
        if (
            actual_txt
            and progress_col
            and metric_txt != "Cost savings target"
            and _promise_value_looks_like_progress(actual_txt, metric=metric_txt)
        ):
            ws.cell(rr, progress_col).value = progress_txt or _promise_progress_label(actual_txt, metric=metric_txt, stated=stated_txt)
            ws.cell(rr, actual_col).value = ""
            if cols.get("status"):
                status_now = str(ws.cell(rr, cols["status"]).value or "").strip().lower()
                if status_now in {"completed", "hit", "missed", "beat"}:
                    ws.cell(rr, cols["status"]).value = "On track"
            actual_txt = ""
            progress_txt = str(ws.cell(rr, progress_col).value or "").strip()
        if not prev_txt and not new_txt and actual_txt and not re.search(r"\b(target|plan|expected|milestone|qualified|operational|run[- ]rate)\b", note_txt, flags=re.I):
            rows_to_delete.add(rr)
            continue
        if new_txt.lower() in {"actual reported", "final actual"} or change_txt == "actual":
            if prev_txt and actual_txt and new_col and change_col:
                ws.cell(rr, new_col).value = prev_txt
                ws.cell(rr, change_col).value = "Maintained"
            else:
                rows_to_delete.add(rr)
                continue
        if "45z monetization" in metric_txt.lower():
            for col in (prev_col, new_col):
                if col:
                    ws.cell(rr, col).value = _clean_gpre_45z_monetization_value(ws.cell(rr, col).value)

    for rr in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(rr, 1)

    # Fill missing previous guide from prior same metric+horizon event and
    # remove duplicate Initial rows inside the same block.
    timeline = _timeline_rows()
    latest_guide: Dict[Tuple[str, str], str] = {}
    for item in sorted(
        timeline,
        key=lambda x: _promise_event_sort_key(
            ws.cell(int(x["row"]), x["cols"].get("stated in", 0)).value if x["cols"].get("stated in") else x["section"],
            ws.cell(int(x["row"]), x["cols"].get("source date", 0)).value if x["cols"].get("source date") else "",
        ),
    ):
        rr = int(item["row"])
        cols = item["cols"]
        metric_col = cols.get("metric")
        horizon_col = cols.get("horizon")
        prev_col = cols.get("previous guide")
        new_col = cols.get("new/current guide")
        change_col = cols.get("change type")
        if not (metric_col and horizon_col and prev_col and new_col and change_col):
            continue
        metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
        horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip()
        key = (_promise_metric_definition_key(metric_txt), horizon_txt.lower())
        prev_txt = str(ws.cell(rr, prev_col).value or "").strip()
        new_txt = str(ws.cell(rr, new_col).value or "").strip()
        if not prev_txt and latest_guide.get(key):
            prev_txt = latest_guide[key]
            ws.cell(rr, prev_col).value = prev_txt
        if prev_txt and new_txt:
            change_txt = str(ws.cell(rr, change_col).value or "").strip().lower()
            if change_txt in {"", "initial"}:
                ws.cell(rr, change_col).value = "Maintained" if prev_txt == new_txt else "Updated"
            elif change_txt == "updated" and prev_txt == new_txt:
                actual_txt = str(ws.cell(rr, cols.get("actual", 0)).value or "").strip() if cols.get("actual") else ""
                progress_txt = str(ws.cell(rr, cols.get("progress / run-rate", 0)).value or "").strip() if cols.get("progress / run-rate") else ""
                if "45z monetization" in metric_txt.lower() and (actual_txt or progress_txt):
                    ws.cell(rr, change_col).value = "Updated"
                else:
                    ws.cell(rr, change_col).value = "Maintained"
        elif new_txt and not prev_txt and not str(ws.cell(rr, change_col).value or "").strip():
            ws.cell(rr, change_col).value = "Initial"
        if new_txt and not re.search(r"\bactual reported|final actual\b", new_txt, flags=re.I):
            latest_guide[key] = new_txt

    grouped: Dict[Tuple[str, str, str], List[int]] = {}
    for item in _timeline_rows():
        rr = int(item["row"])
        cols = item["cols"]
        metric = _promise_metric_definition_key(ws.cell(rr, cols.get("metric", 1)).value if cols.get("metric") else "")
        horizon = str(ws.cell(rr, cols.get("horizon", 0)).value or "").strip().lower() if cols.get("horizon") else ""
        grouped.setdefault((str(item["section"]).lower(), metric, horizon), []).append(rr)

    def _row_keep_score(row_idx: int) -> Tuple[int, int, int, int, int]:
        cols = next((x["cols"] for x in _timeline_rows() if int(x["row"]) == row_idx), {})
        actual = str(ws.cell(row_idx, cols.get("actual", 0)).value or "").strip() if cols.get("actual") else ""
        prev = str(ws.cell(row_idx, cols.get("previous guide", 0)).value or "").strip() if cols.get("previous guide") else ""
        new = str(ws.cell(row_idx, cols.get("new/current guide", 0)).value or "").strip() if cols.get("new/current guide") else ""
        change = str(ws.cell(row_idx, cols.get("change type", 0)).value or "").strip().lower() if cols.get("change type") else ""
        source = str(ws.cell(row_idx, cols.get("source date", 0)).value or "").strip() if cols.get("source date") else ""
        try:
            source_ord = pd.Timestamp(source).date().toordinal()
        except Exception:
            source_ord = 0
        return (
            1 if actual else 0,
            1 if change not in {"", "initial"} else 0,
            1 if prev else 0,
            1 if new else 0,
            source_ord,
        )

    delete_dupes: Set[int] = set()
    for row_nums in grouped.values():
        if len(row_nums) <= 1:
            continue
        keep = sorted(row_nums, key=_row_keep_score, reverse=True)[0]
        keep_cols = next((x["cols"] for x in _timeline_rows() if int(x["row"]) == keep), {})
        for row_idx in row_nums:
            if row_idx == keep:
                continue
            dup_cols = next((x["cols"] for x in _timeline_rows() if int(x["row"]) == row_idx), {})
            for header in ("previous guide", "new/current guide", "actual", "progress / run-rate", "source / note"):
                kc = keep_cols.get(header)
                dc = dup_cols.get(header)
                if kc and dc and not str(ws.cell(keep, kc).value or "").strip() and str(ws.cell(row_idx, dc).value or "").strip():
                    ws.cell(keep, kc).value = ws.cell(row_idx, dc).value
            delete_dupes.add(row_idx)
    for rr in sorted(delete_dupes, reverse=True):
        ws.delete_rows(rr, 1)

    # Keep rows inside each event block in a stable business-metric order after
    # all semantic repairs. This changes presentation only; section/horizon
    # assignment and row formulas stay untouched.
    by_section: Dict[str, List[Dict[str, Any]]] = {}
    for item in _timeline_rows():
        by_section.setdefault(str(item["section"]), []).append(item)
    for section, items in by_section.items():
        if len(items) <= 1:
            continue
        row_nums = [int(item["row"]) for item in sorted(items, key=lambda x: int(x["row"]))]
        cols = items[0]["cols"]
        metric_col = cols.get("metric", 1)
        horizon_col = cols.get("horizon", 0)
        source_col = cols.get("source date", 0)
        stated_col = cols.get("stated in", 0)
        snapshots = [
            [ws.cell(row_idx, cc).value for cc in range(1, max_col + 1)]
            for row_idx in row_nums
        ]

        def _snapshot_sort_key(values: List[Any]) -> Tuple[int, str, str, str, str]:
            metric = values[metric_col - 1] if metric_col else ""
            horizon = values[horizon_col - 1] if horizon_col else ""
            stated = values[stated_col - 1] if stated_col else section
            source = values[source_col - 1] if source_col else ""
            return (
                _promise_metric_order_rank(metric),
                str(metric or "").strip().lower(),
                str(horizon or "").strip().lower(),
                str(stated or "").strip().lower(),
                str(source or "").strip().lower(),
            )

        for row_idx in row_nums:
            for merge_range in list(ws.merged_cells.ranges):
                if merge_range.min_row <= row_idx <= merge_range.max_row:
                    try:
                        ws.unmerge_cells(str(merge_range))
                    except KeyError:
                        pass
        for row_idx, values in zip(row_nums, sorted(snapshots, key=_snapshot_sort_key)):
            for cc, value in enumerate(values, start=1):
                if ws.cell(row_idx, cc).__class__.__name__ == "MergedCell":
                    for merge_range in list(ws.merged_cells.ranges):
                        if (
                            merge_range.min_row <= row_idx <= merge_range.max_row
                            and merge_range.min_col <= cc <= merge_range.max_col
                        ):
                            try:
                                ws.unmerge_cells(str(merge_range))
                            except KeyError:
                                pass
                            break
                cell = ws.cell(row_idx, cc)
                if cell.__class__.__name__ == "MergedCell":
                    if value in (None, ""):
                        continue
                    for merge_range in list(ws.merged_cells.ranges):
                        if (
                            merge_range.min_row <= row_idx <= merge_range.max_row
                            and merge_range.min_col <= cc <= merge_range.max_col
                        ):
                            ws.cell(row_idx, merge_range.min_col).value = value
                            break
                    continue
                cell.value = value

    for item in _timeline_rows():
        if "pre-release" not in str(item.get("section") or "").lower():
            continue
        status_col = item["cols"].get("status")
        if status_col:
            ws.cell(int(item["row"]), status_col).value = "On track"
    _remove_empty_promise_revision_blocks(ws)
    _remove_promise_metric_stubs(ws)

def _promise_status_fill_for_label(value: Any) -> PatternFill:
    low = str(value or "").strip().lower()
    if low in {"completed", "complete", "delivered", "achieved"}:
        return PatternFill("solid", fgColor="009E73")
    if low in {"beat", "hit", "met"}:
        return PatternFill("solid", fgColor="66C2A5")
    if low in {"on track", "on_track"}:
        return PatternFill("solid", fgColor="56B4E9")
    if low in {"open"}:
        return PatternFill("solid", fgColor="A6CEE3")
    if low in {"mixed", "partial"}:
        return PatternFill("solid", fgColor="E69F00")
    if low in {"basis-dependent", "basis dependent"}:
        return PatternFill("solid", fgColor="CC79A7")
    if low in {"missed", "miss", "failed", "fail"}:
        return PatternFill("solid", fgColor="D55E00")
    return PatternFill("solid", fgColor="FFFFFF")

def _apply_promise_grid_style(ws: Any) -> None:
    """Make Promise_Progress_UI visually continuous across visible columns after all edits."""
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    _repair_anf_promise_actual_progress_semantics(ws)
    max_col = PROMISE_VISIBLE_MAX_COL
    section_fill = PatternFill("solid", fgColor="5B9BD5")
    header_fill = PatternFill("solid", fgColor="EAF3FB")
    body_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="F6F9FC")
    border = Border(bottom=Side(style="thin", color="D9E2EF"))

    def _row_values(row_idx: int) -> List[str]:
        return [str(ws.cell(row_idx, cc).value or "").strip() for cc in range(1, max_col + 1)]

    def _is_header(row_idx: int) -> bool:
        vals = {v.lower() for v in _row_values(row_idx) if v}
        return ("metric" in vals or "milestone" in vals or "category" in vals) and (
            "status" in vals or "score" in vals
        )

    active_status_col = 0
    body_counter = 0
    for rr in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        is_section = _is_promise_section_row(ws, rr)
        is_header = _is_header(rr)
        is_blank = all(not v for v in _row_values(rr))
        if is_section:
            for merge_range in list(ws.merged_cells.ranges):
                if merge_range.min_row == rr and merge_range.max_row == rr:
                    try:
                        ws.unmerge_cells(str(merge_range))
                    except KeyError:
                        pass
            for cc in range(1, max_col + 1):
                cell = ws.cell(rr, cc)
                cell.fill = section_fill
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=max_col)
            ws.row_dimensions[rr].height = 24 if first_txt == "Promise Progress" else 22
            active_status_col = 0
            body_counter = 0
            continue
        if is_header:
            for merge_range in list(ws.merged_cells.ranges):
                if merge_range.min_row == rr and merge_range.max_row == rr:
                    try:
                        ws.unmerge_cells(str(merge_range))
                    except KeyError:
                        pass
            vals = _row_values(rr)
            active_status_col = next((idx + 1 for idx, v in enumerate(vals) if v.lower() == "status"), 0)
            for cc in range(1, max_col + 1):
                cell = ws.cell(rr, cc)
                cell.fill = header_fill
                cell.font = Font(bold=True, size=11, color="000000")
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            note_labels = {"notes/source", "source / note"}
            if note_labels & {v.lower() for v in vals}:
                note_col = next((idx + 1 for idx, v in enumerate(vals) if v.lower() in note_labels), 0)
                if note_col and note_col < max_col:
                    ws.merge_cells(start_row=rr, start_column=note_col, end_row=rr, end_column=max_col)
            ws.row_dimensions[rr].height = 22
            body_counter = 0
            continue
        if is_blank:
            for cc in range(1, max_col + 1):
                cell = ws.cell(rr, cc)
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
                cell.border = Border()
            ws.row_dimensions[rr].height = 18
            continue
        body_counter += 1
        fill = alt_fill if body_counter % 2 else body_fill
        for cc in range(1, max_col + 1):
            cell = ws.cell(rr, cc)
            cell.fill = _promise_status_fill_for_label(cell.value) if active_status_col and cc == active_status_col else copy(fill)
            cell.border = border
            cell.font = Font(size=11, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc in {11, max_col})
        ws.row_dimensions[rr].height = 24

    for col, width in {
        "A": 28,
        "B": 28,
        "C": 32,
        "D": 15,
        "E": 22,
        "F": 28,
        "G": 15,
        "H": 14,
        "I": 16,
        "J": 14,
        "K": 42,
        "L": 42,
    }.items():
        ws.column_dimensions[col].width = width
    for col in ("M", "N", "O"):
        ws.column_dimensions[col].hidden = True

def _cleanup_anf_promise_after_repair(ws: Any) -> None:
    """Remove generic annual-helper rows from ANF's curated timeline."""
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    first_col_values = {
        str(ws.cell(rr, 1).value or "").strip()
        for rr in range(1, min(int(ws.max_row or 0), 140) + 1)
    }
    if not ({"Tariff impact", "Real estate activity"} & first_col_values):
        return
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)
    current_section = ""
    active_cols: Dict[str, int] = {}
    rows_to_delete: List[int] = []
    final_actuals: Dict[Tuple[str, str], str] = {}
    scan_section = ""
    scan_cols: Dict[str, int] = {}
    for rr in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        if _is_promise_section_row(ws, rr):
            scan_section = first_txt
            scan_cols = {}
            continue
        headers = {
            str(ws.cell(rr, cc).value or "").strip().lower(): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(rr, cc).value or "").strip()
        }
        if "metric" in headers and "horizon" in headers:
            scan_cols = headers
            continue
        if scan_section != "2025-Q4 revisions" or not scan_cols:
            continue
        metric_col = scan_cols.get("metric")
        horizon_col = scan_cols.get("horizon")
        actual_col = scan_cols.get("actual")
        if not (metric_col and horizon_col and actual_col):
            continue
        metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
        horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip()
        actual_txt = str(ws.cell(rr, actual_col).value or "").strip()
        if metric_txt and horizon_txt and actual_txt:
            key = (_promise_metric_definition_key(metric_txt), horizon_txt.lower())
            final_actuals[key] = actual_txt
    for rr in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        if _is_promise_section_row(ws, rr):
            current_section = first_txt
            active_cols = {}
            continue
        headers = {
            str(ws.cell(rr, cc).value or "").strip().lower(): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(rr, cc).value or "").strip()
        }
        if "metric" in headers and "horizon" in headers:
            active_cols = headers
            continue
        if not active_cols or not current_section.endswith("revisions"):
            continue
        metric_col = active_cols.get("metric")
        actual_col = active_cols.get("actual")
        status_col = active_cols.get("status")
        note_col = active_cols.get("source / note")
        if not metric_col or not str(ws.cell(rr, metric_col).value or "").strip():
            continue
        if "pre-release" in current_section.lower() and actual_col and status_col:
            horizon_col = active_cols.get("horizon")
            metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
            horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip() if horizon_col else ""
            final_key = (_promise_metric_definition_key(metric_txt), horizon_txt.lower())
            final_actual = final_actuals.get(final_key, "")
            if final_actual:
                ws.cell(rr, actual_col).value = final_actual
                if note_col:
                    note_txt = str(ws.cell(rr, note_col).value or "").strip()
                    timing_note = "Year result shown for comparison; pre-release was issued before final report."
                    if timing_note not in note_txt:
                        ws.cell(rr, note_col).value = f"{note_txt} {timing_note}".strip()
            else:
                ws.cell(rr, actual_col).value = ""
            ws.cell(rr, status_col).value = "On track"
        note_txt = str(ws.cell(rr, note_col).value or "").strip() if note_col else ""
        if note_txt == "Tracking against annual guide.":
            rows_to_delete.append(rr)
    for rr in sorted(set(rows_to_delete), reverse=True):
        ws.delete_rows(rr, 1)

def _repair_anf_promise_actual_progress_semantics(ws: Any) -> None:
    """Apply ANF-only source-backed Actual vs Progress semantics.

    ANF Q4 releases report both quarter results and full-year/YTD results. The
    revision log should keep those bases separate: quarter values in Actual and
    full-year/YTD values in Progress / run-rate.
    """
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)
    header_blob = " ".join(
        str(ws.cell(rr, cc).value or "")
        for rr in range(1, min(int(ws.max_row or 0), 8) + 1)
        for cc in range(1, max_col + 1)
    )
    first_col_values = {
        str(ws.cell(rr, 1).value or "").strip()
        for rr in range(1, min(int(ws.max_row or 0), 140) + 1)
    }
    if "ANF guidance tracker" not in header_blob and not ({"Tariff impact", "Real estate activity"} & first_col_values):
        return

    q4_semantics = {
        "Net sales growth": ("+5.4%", "FY: +6%"),
        "Operating margin": ("14.1%", "FY: 13.3% GAAP / 12.5% adjusted"),
        "Adjusted EPS": ("$3.68 adjusted", "FY: $9.86 adjusted"),
        "Capex": ("$55.6m", "FY: $240.8m"),
        "Diluted shares": ("46.8m diluted", "Δ vs guide: -1.2m; Δ YTD: -5.6m"),
        "Share repurchases": ("$100.0m", "FY: $450m"),
    }
    interim_adjusted_eps = {
        "2025-Q1": ("$1.59 adjusted", "YTD: $1.59 adjusted", "Quarter/YTD adjusted EPS; annual guide still open."),
        "2025-Q2": ("$2.32 adjusted", "YTD: $3.91 adjusted", "Quarter/YTD adjusted EPS; annual guide still open."),
        "2025-Q3": ("$2.36 adjusted", "YTD: $6.27 adjusted", "Quarter/YTD adjusted EPS; annual guide still open."),
    }
    interim_diluted_shares = {
        "2025-Q1": ("50.6m diluted", "Δ vs guide: +1.6m; Δ YTD: -1.8m", "Q1 diluted shares; Progress shows share-count delta versus guide and fiscal-year start."),
        "2025-Q2": ("48.6m diluted", "Δ vs guide: -0.4m; Δ YTD: -3.9m", "Q2 diluted shares; Progress shows share-count delta versus guide and fiscal-year start."),
        "2025-Q3": ("47.9m diluted", "Δ vs guide: -0.1m; Δ YTD: -4.6m", "Q3 diluted shares; Progress shows share-count delta versus guide and fiscal-year start."),
    }
    q4_note = "Q4 actual shown in Actual; FY/YTD result shown in Progress / run-rate."
    q4_adjusted_eps_note = "Q4 actual shown in Actual (adjusted EPS); official FY adjusted EPS shown in Progress / run-rate. FY can differ from summed rounded quarters."
    q4_share_note = "Q4 actual shown in Actual (diluted shares); Progress shows share-count reduction versus guide and YTD."
    pre_release_note = "Final Q4/FY results shown for comparison; pre-release was issued before final report."
    final_q4_rows = {
        "Net sales growth": ("at least +6%", "at least +6%", "Completed"),
        "Operating margin": ("around 13%", "around 13%", "Mixed"),
        "Adjusted EPS": ("$10.30-$10.40", "$10.30-$10.40", "Missed"),
        "Capex": ("~$245m", "~$245m", "Hit"),
        "Diluted shares": ("~48m", "~48m", "Completed"),
        "Share repurchases": ("~$450m", "~$450m", "Completed"),
    }

    def _ensure_final_q4_revision_block() -> None:
        def _write_final_rows(row_idx: int, missing_metrics: Iterable[str]) -> None:
            for metric_txt in missing_metrics:
                for merged in list(ws.merged_cells.ranges):
                    if (
                        int(getattr(merged, "min_row", 0)) == row_idx
                        and int(getattr(merged, "max_row", 0)) == row_idx
                        and int(getattr(merged, "min_col", 0)) <= max_col
                    ):
                        try:
                            ws.unmerge_cells(str(merged))
                        except KeyError:
                            try:
                                ws.merged_cells.ranges.remove(merged)
                            except (KeyError, ValueError):
                                pass
                            for cc in range(int(merged.min_col) + 1, int(merged.max_col) + 1):
                                cell_key = (row_idx, cc)
                                if ws._cells.get(cell_key).__class__.__name__ == "MergedCell":
                                    del ws._cells[cell_key]
                prev_txt, new_txt, status_txt = final_q4_rows[metric_txt]
                actual_txt, progress_txt = q4_semantics[metric_txt]
                row_values = [
                    metric_txt,
                    prev_txt,
                    new_txt,
                    "Completed",
                    actual_txt,
                    progress_txt,
                    status_txt,
                    "2025 year",
                    "2025-Q4",
                    "2026-03-04",
                    q4_note,
                ]
                for cc, value in enumerate(row_values, start=1):
                    ws.cell(row_idx, cc).value = value
                row_idx += 1

        for row_idx in range(1, int(ws.max_row or 0) + 1):
            if str(ws.cell(row_idx, 1).value or "").strip() == "2025-Q4 revisions":
                header_row = row_idx + 1
                existing_metrics: set[str] = set()
                block_end = int(ws.max_row or 0) + 1
                for scan_row in range(header_row + 1, int(ws.max_row or 0) + 1):
                    first_txt = str(ws.cell(scan_row, 1).value or "").strip()
                    if first_txt.endswith("revisions"):
                        block_end = scan_row
                        break
                    if first_txt:
                        existing_metrics.add(first_txt)
                missing_metrics = [metric for metric in final_q4_rows if metric not in existing_metrics]
                if missing_metrics:
                    ws.insert_rows(block_end, len(missing_metrics))
                    _write_final_rows(block_end, missing_metrics)
                return
        insert_at = int(ws.max_row or 0) + 1
        pre_release_row = 0
        for row_idx in range(1, int(ws.max_row or 0) + 1):
            if str(ws.cell(row_idx, 1).value or "").strip() == "2025-Q4 pre-release update revisions":
                pre_release_row = row_idx
                continue
            if pre_release_row and row_idx > pre_release_row and str(ws.cell(row_idx, 1).value or "").strip().endswith("revisions"):
                insert_at = row_idx
                break
        blue = PatternFill("solid", fgColor="5B9BD5")
        header_fill = PatternFill("solid", fgColor="EAF3FB")
        ws.insert_rows(insert_at, 2 + len(final_q4_rows))
        ws.cell(insert_at, 1).value = "2025-Q4 revisions"
        for cc in range(1, max_col + 1):
            ws.cell(insert_at, cc).fill = blue
        for cc, header in enumerate(PROMISE_TIMELINE_HEADERS, start=1):
            ws.cell(insert_at + 1, cc).value = header
            ws.cell(insert_at + 1, cc).fill = header_fill
        _write_final_rows(insert_at + 2, final_q4_rows.keys())

    _ensure_final_q4_revision_block()

    current_section = ""
    active_cols: Dict[str, int] = {}
    for rr in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        if _is_promise_section_row(ws, rr):
            current_section = first_txt
            active_cols = {}
            continue
        row_map = {
            _promise_header_name(ws.cell(rr, cc).value): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(rr, cc).value or "").strip()
        }
        if "metric" in row_map and "actual" in row_map:
            active_cols = row_map
            continue
        if not active_cols or not current_section.endswith("revisions"):
            continue

        metric_col = active_cols.get("metric")
        actual_col = active_cols.get("actual")
        progress_col = active_cols.get("progress / run-rate")
        status_col = active_cols.get("status")
        horizon_col = active_cols.get("horizon")
        stated_col = active_cols.get("stated in")
        note_col = active_cols.get("source / note") or active_cols.get("notes/source")
        if not (metric_col and actual_col and horizon_col):
            continue
        metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
        if not metric_txt:
            continue
        horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip()
        stated_txt = str(ws.cell(rr, stated_col).value or "").strip() if stated_col else ""

        if metric_txt == "Adjusted EPS" and horizon_txt == "2025 year" and stated_txt in interim_adjusted_eps:
            actual_txt, progress_txt, note_txt = interim_adjusted_eps[stated_txt]
            _set_promise_row_semantics(
                ws,
                rr,
                active_cols,
                actual=actual_txt,
                progress=progress_txt,
                status="On track",
                note=note_txt,
            )
            continue

        if metric_txt == "Diluted shares" and horizon_txt == "2025 year" and stated_txt in interim_diluted_shares:
            actual_txt, progress_txt, note_txt = interim_diluted_shares[stated_txt]
            _set_promise_row_semantics(
                ws,
                rr,
                active_cols,
                actual=actual_txt,
                progress=progress_txt,
                status="On track",
                note=note_txt,
            )
            continue

        if current_section not in {"2025-Q4 revisions", "2025-Q4 pre-release update revisions"}:
            continue
        if horizon_txt != "2025 year" or metric_txt not in q4_semantics:
            continue
        actual_txt, progress_txt = q4_semantics[metric_txt]
        row_note = pre_release_note if "pre-release" in current_section.lower() else q4_note
        if metric_txt == "Adjusted EPS":
            row_note = pre_release_note if "pre-release" in current_section.lower() else q4_adjusted_eps_note
        elif metric_txt == "Diluted shares":
            row_note = q4_share_note
        _set_promise_row_semantics(
            ws,
            rr,
            active_cols,
            change_type="Completed" if current_section == "2025-Q4 revisions" else None,
            actual=actual_txt,
            progress=progress_txt,
            status="On track" if "pre-release" in current_section.lower() else None,
            note=row_note,
        )
    _ensure_anf_promise_hidden_source_keys(ws)

def _clear_pre_release_promise_actuals(ws: Any) -> None:
    """Pre-release blocks are guidance revisions, not final-result blocks."""
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    wb_obj = getattr(ws, "parent", None)
    if wb_obj is not None and any(str(name).startswith("ANF_") for name in getattr(wb_obj, "sheetnames", [])):
        return
    header_blob = " ".join(str(ws.cell(rr, cc).value or "") for rr in range(1, min(int(ws.max_row or 0), 3) + 1) for cc in range(1, PROMISE_VISIBLE_MAX_COL + 1))
    if "ANF guidance tracker" in header_blob:
        return
    early_values = {
        str(ws.cell(rr, 1).value or "").strip()
        for rr in range(1, min(int(ws.max_row or 0), 80) + 1)
    }
    if {"Tariff impact", "Real estate activity"} & early_values:
        return
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)
    current_section = ""
    active_cols: Dict[str, int] = {}
    for rr in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        if _is_promise_section_row(ws, rr):
            current_section = first_txt
            active_cols = {}
            continue
        headers = {
            str(ws.cell(rr, cc).value or "").strip().lower(): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(rr, cc).value or "").strip()
        }
        if "metric" in headers and "actual" in headers:
            active_cols = headers
            continue
        if "pre-release" not in current_section.lower() or not active_cols:
            continue
        metric_col = active_cols.get("metric")
        actual_col = active_cols.get("actual")
        status_col = active_cols.get("status")
        if not metric_col or not actual_col or not status_col:
            continue
        if not str(ws.cell(rr, metric_col).value or "").strip():
            continue
        ws.cell(rr, actual_col).value = ""
        ws.cell(rr, status_col).value = "On track"

def _remove_empty_promise_revision_blocks(ws: Any) -> None:
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    max_row = int(ws.max_row or 0)
    if max_row < 3:
        return
    section_rows: List[int] = []
    for rr in range(1, max_row + 1):
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        first_fill = str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper()
        if first_txt.endswith("revisions") and first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")):
            section_rows.append(rr)
    delete_ranges: List[Tuple[int, int]] = []
    for idx, start_row in enumerate(section_rows):
        next_section = section_rows[idx + 1] if idx + 1 < len(section_rows) else max_row + 1
        end_row = next_section - 1
        has_data = False
        for rr in range(start_row + 1, end_row + 1):
            first_txt = str(ws.cell(rr, 1).value or "").strip()
            if not first_txt or first_txt.lower() in {"metric", "milestone", "category"}:
                continue
            other_values = [
                str(ws.cell(rr, cc).value or "").strip()
                for cc in range(2, min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL) + 1)
            ]
            if any(other_values):
                has_data = True
                break
        if not has_data:
            delete_ranges.append((start_row, end_row))
    for start_row, end_row in sorted(delete_ranges, reverse=True):
        ws.delete_rows(start_row, max(1, end_row - start_row + 1))

def _standardize_promise_section_layout(ws: Any) -> None:
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    max_col = PROMISE_VISIBLE_MAX_COL
    suffixes = ("5B9BD5", "6FA8DC", "4472C4")
    for row_idx in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(row_idx, 1).value or "").strip()
        first_fill = str(ws.cell(row_idx, 1).fill.fgColor.rgb or "").upper()
        if not first_txt or not first_fill.endswith(suffixes):
            continue
        for merge_range in list(ws.merged_cells.ranges):
            if merge_range.min_row == row_idx and merge_range.max_row == row_idx:
                ws.unmerge_cells(str(merge_range))
        fill = copy(ws.cell(row_idx, 1).fill)
        font = copy(ws.cell(row_idx, 1).font)
        border = copy(ws.cell(row_idx, 1).border)
        alignment = copy(ws.cell(row_idx, 1).alignment)
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = copy(fill)
            cell.font = copy(font)
            cell.border = copy(border)
            cell.alignment = copy(alignment)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
        ws.row_dimensions[row_idx].height = 26 if first_txt == "Promise Progress" else 22
    for col in ("M", "N"):
        ws.column_dimensions[col].width = 4
        ws.column_dimensions[col].hidden = True
    ws.column_dimensions["O"].hidden = True
    for col, width in {
        "A": 28,
        "B": 28,
        "C": 32,
        "D": 15,
        "E": 22,
        "F": 28,
        "G": 15,
        "H": 14,
        "I": 16,
        "J": 14,
        "K": 42,
        "L": 42,
    }.items():
        ws.column_dimensions[col].width = max(float(ws.column_dimensions[col].width or 0), float(width))

def _is_promise_section_row(ws: Any, row_idx: int) -> bool:
    first_txt = str(ws.cell(row_idx, 1).value or "").strip()
    if not first_txt:
        return False
    first_fill = str(ws.cell(row_idx, 1).fill.fgColor.rgb or "").upper()
    return first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")) or first_txt in {
        "Promise Progress",
        "Management Credibility Scorecard",
        "Quarterly guidance timeline / revision log",
    } or first_txt.endswith(
        ("progression", "open guidance", "revisions")
    )

def _ensure_q4_annual_actual_revision_rows(ws: Any) -> None:
    """Add compact Q4 annual rows when actuals exist only in progression.

    The row is not an "Actual reported" special row: it keeps the latest
    guide in Previous/New and fills the Actual column for the matching annual
    horizon. This gives Q4 blocks a proper actual cell without duplicating the
    progression table semantics.
    """
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    wb_obj = getattr(ws, "parent", None)
    if wb_obj is not None and any(str(name).startswith("ANF_") for name in getattr(wb_obj, "sheetnames", [])):
        return
    header_blob = " ".join(
        str(ws.cell(rr, cc).value or "")
        for rr in range(1, min(int(ws.max_row or 0), 25) + 1)
        for cc in range(1, min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL) + 1)
    )
    if "ANF guidance tracker" in header_blob or "Tariff impact" in header_blob or "Real estate activity" in header_blob:
        return
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)

    def _norm(value: Any) -> str:
        return str(value or "").strip()

    def _section_score(title: Any) -> int:
        m = re.search(r"\b(20\d{2})-Q([1-4])\b", str(title or ""), flags=re.I)
        return int(m.group(1)) * 10 + int(m.group(2)) if m else 0

    def _actual_lookup() -> Dict[str, Dict[str, float]]:
        wb = getattr(ws, "parent", None)
        out: Dict[str, Dict[str, float]] = {}
        if wb is None:
            return out

        def _hdr(sheet: Any) -> Dict[str, int]:
            return {
                re.sub(r"[^a-z0-9]+", "_", str(sheet.cell(1, cc).value or "").strip().lower()).strip("_"): cc
                for cc in range(1, int(sheet.max_column or 0) + 1)
            }

        def _num(v: Any) -> Optional[float]:
            val = pd.to_numeric(v, errors="coerce")
            if pd.isna(val):
                return None
            f = float(val)
            return f if math.isfinite(f) else None

        def _labels(qd: date, row_map: Mapping[str, Any]) -> List[str]:
            labels = {str(qd)}
            fiscal_label = str(row_map.get("fiscal_label") or "").strip()
            fy = _num(row_map.get("fiscal_year"))
            fq = _num(row_map.get("fiscal_quarter"))
            if fiscal_label:
                labels.add(fiscal_label)
            elif fy is not None and fq is not None and 1 <= int(fq) <= 4:
                labels.add(f"{int(fy)}-Q{int(fq)}")
            else:
                labels.add(f"{qd.year}-Q{((qd.month - 1) // 3) + 1}")
            return [label for label in labels if label]

        def _add(labels: Iterable[str], key: str, value: Optional[float]) -> None:
            if value is None:
                return
            for label in labels:
                out.setdefault(label, {})[key] = value

        if "History_Q" in getattr(wb, "sheetnames", []):
            sh = wb["History_Q"]
            headers = _hdr(sh)
            q_col = headers.get("quarter")
            if q_col:
                for row_idx in range(2, int(sh.max_row or 0) + 1):
                    qd = _date_or_none(sh.cell(row_idx, q_col).value)
                    if qd is None:
                        continue
                    row_map = {name: sh.cell(row_idx, col_idx).value for name, col_idx in headers.items()}
                    labels = _labels(qd, row_map)
                    rev = _num(row_map.get("revenue"))
                    op = _num(row_map.get("op_income"))
                    cfo = _num(row_map.get("cfo"))
                    capex = _num(row_map.get("capex"))
                    _add(labels, "revenue", rev)
                    _add(labels, "operating_margin", (op / rev) if rev and op is not None else None)
                    _add(labels, "fcf", (cfo - capex) if cfo is not None and capex is not None else None)
                    _add(labels, "capex", capex)
                    _add(labels, "eps", _num(row_map.get("eps_diluted")))
                    _add(labels, "shares", _num(row_map.get("shares_diluted")))
                    _add(labels, "buybacks", _num(row_map.get("buybacks_cash")))
        if "Adjusted_Metrics" in getattr(wb, "sheetnames", []):
            sh = wb["Adjusted_Metrics"]
            headers = _hdr(sh)
            q_col = headers.get("quarter")
            if q_col:
                for row_idx in range(2, int(sh.max_row or 0) + 1):
                    qd = _date_or_none(sh.cell(row_idx, q_col).value)
                    if qd is None:
                        continue
                    row_map = {name: sh.cell(row_idx, col_idx).value for name, col_idx in headers.items()}
                    labels = _labels(qd, row_map)
                    _add(labels, "adj_ebit", _num(row_map.get("adj_ebit")))
                    _add(labels, "adj_ebitda", _num(row_map.get("adj_ebitda")))
                    _add(labels, "adj_eps", _num(row_map.get("adj_eps")))
                    _add(labels, "adj_fcf", _num(row_map.get("adj_fcf")))
        return out

    period_actuals = _actual_lookup()

    def _actual_key(metric: Any) -> str:
        low = str(metric or "").lower()
        if "adjusted ebitda" in low or "adj ebitda" in low:
            return "adj_ebitda"
        if "adjusted ebit" in low or "adj ebit" in low:
            return "adj_ebit"
        if "fcf" in low or "free cash" in low:
            return "fcf"
        if "eps" in low:
            return "adj_eps" if "adjusted" in low or "adj" in low else "eps"
        if "operating margin" in low:
            return "operating_margin"
        if "capex" in low:
            return "capex"
        if "repurchase" in low or "buyback" in low:
            return "buybacks"
        if "share" in low:
            return "shares"
        if "revenue" in low or "sales" in low:
            return "revenue"
        return ""

    def _fmt_actual(metric: Any, key: str, value: Any) -> str:
        val = pd.to_numeric(value, errors="coerce")
        if pd.isna(val):
            return ""
        num = float(val)
        if key == "operating_margin":
            return f"{num * 100:.1f}%"
        if key in {"eps", "adj_eps"}:
            return f"${num:,.2f}"
        if key == "shares":
            return f"{num / 1_000_000:,.1f}m" if abs(num) > 1_000_000 else f"{num:,.1f}m"
        if abs(num) >= 1_000_000_000:
            return f"${num / 1_000_000_000:,.2f}bn"
        if abs(num) >= 100_000:
            return f"${num / 1_000_000:,.1f}m"
        return f"${num:,.1f}m"

    def _period_actual(metric: Any, label: str) -> str:
        key = _actual_key(metric)
        if not key:
            return ""
        vals = period_actuals.get(label, {})
        if key in vals:
            return _fmt_actual(metric, key, vals[key])
        if key == "eps" and "adj_eps" in vals:
            return _fmt_actual(metric, "adj_eps", vals["adj_eps"])
        if key == "fcf" and "adj_fcf" in vals:
            return _fmt_actual(metric, "adj_fcf", vals["adj_fcf"])
        return ""

    def _annual_actual(metric: Any, year: int) -> str:
        key = _actual_key(metric)
        if key not in {"revenue", "adj_ebit", "adj_ebitda", "fcf", "adj_fcf", "capex", "buybacks"}:
            return ""
        vals: List[float] = []
        actual_key = key
        for q_num in (1, 2, 3, 4):
            period_vals = period_actuals.get(f"{year}-Q{q_num}", {})
            value = period_vals.get(key)
            if value is None and key == "fcf":
                value = period_vals.get("adj_fcf")
                if value is not None:
                    actual_key = "adj_fcf"
            numeric = pd.to_numeric(value, errors="coerce")
            if pd.isna(numeric):
                return ""
            vals.append(float(numeric))
        return _fmt_actual(metric, actual_key, sum(vals)) if len(vals) == 4 else ""

    def _change_label(prev_in: Any, new_in: Any) -> str:
        prev_txt = _norm(prev_in)
        new_txt = _norm(new_in)
        if not prev_txt:
            return "Initial"
        return "Maintained" if prev_txt == new_txt else "Updated"

    annual_rows: List[Dict[str, str]] = []
    rr = 1
    while rr <= int(ws.max_row or 0):
        title = _norm(ws.cell(rr, 1).value)
        m = re.fullmatch(r"(20\d{2}) guidance progression", title, flags=re.I)
        if not m:
            rr += 1
            continue
        year = int(m.group(1))
        header_row = rr + 1
        headers = {
            _norm(ws.cell(header_row, cc).value).lower(): cc
            for cc in range(1, max_col + 1)
            if _norm(ws.cell(header_row, cc).value)
        }
        metric_col = headers.get("metric")
        actual_col = headers.get("actual")
        status_col = headers.get("status")
        if not (metric_col and actual_col and status_col):
            rr += 1
            continue
        data_row = header_row + 1
        while data_row <= int(ws.max_row or 0) and not _is_promise_section_row(ws, data_row):
            metric = _norm(ws.cell(data_row, metric_col).value)
            actual = _norm(ws.cell(data_row, actual_col).value)
            status = _norm(ws.cell(data_row, status_col).value)
            if metric:
                guide_by_q: Dict[int, str] = {}
                for q_num, label in ((1, "q1 update"), (2, "q2 update"), (3, "q3 update"), (4, "q4 update")):
                    col_idx = headers.get(label)
                    val = _norm(ws.cell(data_row, col_idx).value) if col_idx else ""
                    if val:
                        guide_by_q[q_num] = val
                if actual:
                    guide_by_q.setdefault(4, "")
                latest_guide = ""
                for label in ("q4 update", "jan 2026 update", "q3 update", "q2 update", "q1 update", "initial guide"):
                    col_idx = headers.get(label)
                    candidate = _norm(ws.cell(data_row, col_idx).value) if col_idx else ""
                    if candidate and not re.search(r"\bactual\b", candidate, flags=re.I):
                        latest_guide = candidate
                        break
                if latest_guide:
                    q4_col = headers.get("q4 update")
                    if q4_col and not _norm(ws.cell(data_row, q4_col).value):
                        ws.cell(data_row, q4_col).value = latest_guide
                    guide_by_q[4] = guide_by_q.get(4) or latest_guide
                prior_guide = _norm(ws.cell(data_row, headers.get("initial guide")).value) if headers.get("initial guide") else ""
                for q_num in (1, 2, 3, 4):
                    guide = guide_by_q.get(q_num, "")
                    if not guide:
                        continue
                    period = f"{year}-Q{q_num}"
                    row_actual = (actual or _annual_actual(metric, year)) if q_num == 4 else ""
                    row_progress = "" if q_num == 4 else _promise_progress_label(_period_actual(metric, period), metric=metric, stated=period)
                    row_status = status or ("Completed" if q_num == 4 else "On track")
                    annual_rows.append(
                        {
                            "year": str(year),
                            "metric": metric,
                            "previous": prior_guide,
                            "guide": guide,
                            "actual": row_actual,
                            "progress": row_progress,
                            "status": row_status,
                            "horizon": f"{year} year",
                            "stated": period,
                            "source_date": f"{year}-{q_num * 3:02d}-{'31' if q_num in {1,4} else '30'}",
                            "note": "Reported result for matching annual horizon." if q_num == 4 else "Quarter actual; annual guide still open.",
                        }
                    )
                    prior_guide = guide
            data_row += 1
        rr = data_row

    if not annual_rows:
        return

    # Existing revision keys, plus positions of revision sections.
    existing: Set[Tuple[str, str, str]] = set()
    revision_sections: Dict[str, int] = {}
    active_cols: Dict[str, int] = {}
    current_section = ""
    for row_idx in range(1, int(ws.max_row or 0) + 1):
        first_txt = _norm(ws.cell(row_idx, 1).value)
        if _is_promise_section_row(ws, row_idx):
            current_section = first_txt
            if first_txt.endswith("revisions"):
                revision_sections[first_txt.replace(" revisions", "")] = row_idx
            active_cols = {}
            continue
        headers = {
            _norm(ws.cell(row_idx, cc).value).lower(): cc
            for cc in range(1, max_col + 1)
            if _norm(ws.cell(row_idx, cc).value)
        }
        if "metric" in headers and "horizon" in headers:
            active_cols = headers
            continue
        if not active_cols or "revisions" not in current_section:
            continue
        metric_col = active_cols.get("metric")
        horizon_col = active_cols.get("horizon")
        stated_col = active_cols.get("stated in")
        if metric_col and horizon_col:
            metric = _norm(ws.cell(row_idx, metric_col).value)
            horizon = _norm(ws.cell(row_idx, horizon_col).value)
            stated = _norm(ws.cell(row_idx, stated_col).value) if stated_col else current_section.replace(" revisions", "")
            if metric:
                existing.add((metric.lower(), horizon.lower(), stated.lower()))

    rows_by_block: Dict[str, List[Dict[str, str]]] = {}
    for row in annual_rows:
        key = (row["metric"].lower(), row["horizon"].lower(), row["stated"].lower())
        if key in existing:
            continue
        rows_by_block.setdefault(row["stated"], []).append(row)
    if not rows_by_block:
        return

    blue = PatternFill("solid", fgColor="5B9BD5")
    header_fill = PatternFill("solid", fgColor="EAF3FB")
    neutral = PatternFill("solid", fgColor="FFFFFF")
    neutral_alt = PatternFill("solid", fgColor="F6F9FC")
    border = Border(bottom=Side(style="thin", color="D9E2EF"))

    def _status_fill(status: Any) -> PatternFill:
        low = str(status or "").strip().lower()
        if low in {"completed"}:
            return PatternFill("solid", fgColor="009E73")
        if low in {"beat", "hit"}:
            return PatternFill("solid", fgColor="66C2A5")
        if low in {"on track"}:
            return PatternFill("solid", fgColor="56B4E9")
        if low in {"open"}:
            return PatternFill("solid", fgColor="A6CEE3")
        if low in {"mixed"}:
            return PatternFill("solid", fgColor="E69F00")
        if low in {"basis-dependent"}:
            return PatternFill("solid", fgColor="CC79A7")
        if low in {"missed"}:
            return PatternFill("solid", fgColor="D55E00")
        return neutral

    def _write_section(row_idx: int, block: str) -> None:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=PROMISE_VISIBLE_MAX_COL)
        for cc in range(1, PROMISE_VISIBLE_MAX_COL + 1):
            c = ws.cell(row_idx, cc)
            c.fill = blue
            c.border = border
        c0 = ws.cell(row_idx, 1, f"{block} revisions")
        c0.font = Font(bold=True, size=12, color="FFFFFF")
        c0.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_idx].height = 22

    def _write_header(row_idx: int) -> None:
        labels = PROMISE_TIMELINE_HEADERS
        for cc, label in enumerate(labels, start=1):
            cell = ws.cell(row_idx, cc, label)
            cell.fill = header_fill
            cell.font = Font(bold=True, size=11)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_idx].height = 22

    def _find_insert_row(block: str) -> int:
        target_score = _section_score(block)
        for row_idx in range(1, int(ws.max_row or 0) + 1):
            first_txt = _norm(ws.cell(row_idx, 1).value)
            if first_txt.endswith("revisions") and _section_score(first_txt) < target_score:
                return row_idx
        return int(ws.max_row or 0) + 1

    # Insert from oldest/lower blocks upward so stored row positions for
    # sections above are not invalidated by earlier insertions.
    for block in sorted(rows_by_block, key=_section_score):
        rows = rows_by_block[block]
        if not rows:
            continue
        if block in revision_sections:
            section_row = revision_sections[block]
            insert_at = section_row + 1
            while insert_at <= int(ws.max_row or 0):
                next_first = _norm(ws.cell(insert_at, 1).value)
                if insert_at > section_row + 1 and _is_promise_section_row(ws, insert_at):
                    break
                if any(_norm(ws.cell(insert_at, cc).value) for cc in range(1, PROMISE_VISIBLE_MAX_COL + 1)):
                    insert_at += 1
                    continue
                break
        else:
            insert_at = _find_insert_row(block)
            ws.insert_rows(insert_at, 3)
            _write_section(insert_at, block)
            _write_header(insert_at + 1)
            insert_at += 2
        ws.insert_rows(insert_at, len(rows))
        for offset, row in enumerate(rows):
            row_idx = insert_at + offset
            fill = neutral_alt if row_idx % 2 else neutral
            values = [
                row["metric"],
                row.get("previous") or row["guide"],
                row["guide"],
                _change_label(row.get("previous") or "", row["guide"]),
                row["actual"],
                row.get("progress", ""),
                row["status"],
                row["horizon"],
                row["stated"],
                row["source_date"],
                row["note"],
            ]
            for cc, value in enumerate(values, start=1):
                cell = ws.cell(row_idx, cc, value)
                cell.fill = _status_fill(value) if cc == 7 else fill
                cell.font = Font(size=11)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc == 11)
            ws.row_dimensions[row_idx].height = 24

def _ensure_promise_block_spacing(ws: Any) -> None:
    """Keep one compact blank row between Promise sections/blocks."""
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    row_idx = 2
    while row_idx <= int(ws.max_row or 0):
        first_txt = str(ws.cell(row_idx, 1).value or "").strip()
        if not _is_promise_section_row(ws, row_idx):
            row_idx += 1
            continue
        if first_txt in {"Promise Progress", "Management Credibility Scorecard"}:
            row_idx += 1
            continue
        prev_has_value = any(str(ws.cell(row_idx - 1, cc).value or "").strip() for cc in range(1, PROMISE_VISIBLE_MAX_COL + 1))
        prev_is_section = _is_promise_section_row(ws, row_idx - 1)
        if prev_has_value and not prev_is_section:
            ws.insert_rows(row_idx, 1)
            ws.row_dimensions[row_idx].height = 8.0
            row_idx += 1
        row_idx += 1

def _dedupe_promise_progress_rows(ws: Any) -> None:
    """Merge duplicate Promise rows by visible table key without inventing data."""
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)
    active_cols: Dict[str, int] = {}
    current_section = ""
    seen: Dict[Tuple[str, ...], int] = {}
    rows_to_delete: List[int] = []

    def _norm_header(value: Any) -> str:
        txt = str(value or "").strip().lower()
        if txt in {"actual / latest actual", "latest actual", "latest result"}:
            return "actual"
        return txt

    for rr in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        if _is_promise_section_row(ws, rr):
            current_section = first_txt
            active_cols = {}
            seen = {}
            continue
        row_headers = {
            _norm_header(ws.cell(rr, cc).value): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(rr, cc).value or "").strip()
        }
        if "metric" in row_headers or "milestone" in row_headers or "category" in row_headers:
            active_cols = row_headers
            seen = {}
            continue
        metric_col = active_cols.get("metric") or active_cols.get("milestone")
        if not metric_col:
            continue
        metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
        if not metric_txt:
            continue
        if current_section.endswith("guidance progression"):
            key = (current_section.lower(), metric_txt.lower())
        elif current_section.endswith("open guidance") or current_section == "Open guidance":
            horizon_col = active_cols.get("horizon")
            key = (
                current_section.lower(),
                metric_txt.lower(),
                str(ws.cell(rr, horizon_col).value or "").strip().lower() if horizon_col else "",
            )
        elif current_section.endswith("revisions"):
            horizon_col = active_cols.get("horizon")
            stated_col = active_cols.get("stated in")
            source_col = active_cols.get("source date")
            key = (
                current_section.lower(),
                metric_txt.lower(),
                str(ws.cell(rr, horizon_col).value or "").strip().lower() if horizon_col else "",
                str(ws.cell(rr, stated_col).value or "").strip().lower() if stated_col else "",
                str(ws.cell(rr, source_col).value or "").strip().lower() if source_col else "",
            )
        else:
            key = (current_section.lower(), metric_txt.lower())
        if key not in seen:
            seen[key] = rr
            continue
        keep_rr = seen[key]
        for cc in range(1, max_col + 1):
            keep_val = str(ws.cell(keep_rr, cc).value or "").strip()
            dup_val = str(ws.cell(rr, cc).value or "").strip()
            if not keep_val and dup_val:
                ws.cell(keep_rr, cc).value = ws.cell(rr, cc).value
        rows_to_delete.append(rr)
    for rr in sorted(set(rows_to_delete), reverse=True):
        ws.delete_rows(rr, 1)

def _remove_actual_only_promise_rows(ws: Any) -> None:
    """Remove actual-only timeline rows; actuals belong in the Actual column."""
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)
    active_cols: Dict[str, int] = {}
    rows_to_delete: List[int] = []

    def _norm_header(value: Any) -> str:
        txt = str(value or "").strip().lower()
        if txt in {"actual / latest actual", "latest actual", "latest result"}:
            return "actual"
        return txt

    for rr in range(1, int(ws.max_row or 0) + 1):
        if _is_promise_section_row(ws, rr):
            active_cols = {}
            continue
        row_headers = {
            _norm_header(ws.cell(rr, cc).value): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(rr, cc).value or "").strip()
        }
        if "metric" in row_headers and "change type" in row_headers:
            active_cols = row_headers
            continue
        if not active_cols:
            continue
        metric_col = active_cols.get("metric")
        new_col = active_cols.get("new/current guide")
        change_col = active_cols.get("change type")
        if not metric_col:
            continue
        metric_txt = str(ws.cell(rr, metric_col).value or "").strip().lower()
        new_txt = str(ws.cell(rr, new_col).value or "").strip().lower() if new_col else ""
        change_txt = str(ws.cell(rr, change_col).value or "").strip().lower() if change_col else ""
        if (
            new_txt in {"actual reported", "final actual"}
            or change_txt == "actual"
            or (metric_txt.endswith(" actual") and "guidance" not in metric_txt)
        ):
            rows_to_delete.append(rr)
    for rr in sorted(set(rows_to_delete), reverse=True):
        ws.delete_rows(rr, 1)

def _remove_promise_metric_stubs(ws: Any) -> None:
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)
    active_metric_col: Optional[int] = None
    rows_to_delete: List[int] = []
    for rr in range(1, int(ws.max_row or 0) + 1):
        if _is_promise_section_row(ws, rr):
            active_metric_col = None
            continue
        row_headers = {
            str(ws.cell(rr, cc).value or "").strip().lower(): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(rr, cc).value or "").strip()
        }
        if "metric" in row_headers or "milestone" in row_headers:
            active_metric_col = row_headers.get("metric") or row_headers.get("milestone")
            continue
        if not active_metric_col:
            continue
        metric_txt = str(ws.cell(rr, active_metric_col).value or "").strip()
        if not metric_txt:
            continue
        other_has_value = any(
            str(ws.cell(rr, cc).value or "").strip()
            for cc in range(1, max_col + 1)
            if cc != active_metric_col
        )
        if not other_has_value:
            rows_to_delete.append(rr)
    for rr in sorted(set(rows_to_delete), reverse=True):
        ws.delete_rows(rr, 1)

def _remove_pbi_duplicate_cost_savings_timeline_rows(ws: Any) -> None:
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)
    top_blob = " ".join(
        str(ws.cell(rr, cc).value or "")
        for rr in range(1, min(int(ws.max_row or 0), 4) + 1)
        for cc in range(1, max_col + 1)
    ).lower()
    if "pbi guidance dashboard" not in top_blob:
        return

    rows_to_delete: List[int] = []
    current_section = ""
    active_cols: Dict[str, int] = {}
    cost_rows_by_section: Dict[str, List[int]] = {}
    cost_row_meta: Dict[int, Dict[str, str]] = {}
    for rr in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        if _is_promise_section_row(ws, rr):
            current_section = first_txt
            active_cols = {}
            continue
        row_headers = {
            _promise_header_name(ws.cell(rr, cc).value): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(rr, cc).value or "").strip()
        }
        if "metric" in row_headers and ("new/current guide" in row_headers or "previous guide" in row_headers):
            active_cols = row_headers
            continue
        if not current_section.endswith("revisions") or not active_cols:
            continue
        metric_col = active_cols.get("metric")
        if not metric_col or str(ws.cell(rr, metric_col).value or "").strip() != "Cost savings target":
            continue
        row_meta = {
            "horizon": str(ws.cell(rr, active_cols.get("horizon", 0)).value or "").strip() if active_cols.get("horizon") else "",
            "stated": str(ws.cell(rr, active_cols.get("stated in", 0)).value or "").strip() if active_cols.get("stated in") else "",
            "actual": str(ws.cell(rr, active_cols.get("actual", 0)).value or "").strip() if active_cols.get("actual") else "",
            "progress": str(ws.cell(rr, active_cols.get("progress / run-rate", 0)).value or "").strip() if active_cols.get("progress / run-rate") else "",
            "note": str(ws.cell(rr, active_cols.get("source / note", 0)).value or "").strip() if active_cols.get("source / note") else "",
        }
        cost_rows_by_section.setdefault(current_section, []).append(rr)
        cost_row_meta[rr] = row_meta

    for section, row_nums in cost_rows_by_section.items():
        annualized_rows = [
            rr
            for rr in row_nums
            if cost_row_meta.get(rr, {}).get("horizon", "").lower() == "annualized program"
        ]
        if not annualized_rows:
            continue
        annualized_stated = {cost_row_meta.get(rr, {}).get("stated", "") for rr in annualized_rows}
        for rr in row_nums:
            if rr in annualized_rows:
                continue
            meta = cost_row_meta.get(rr, {})
            horizon = meta.get("horizon", "")
            blob = " ".join(
                str(meta.get(field, "") or "")
                for field in ("actual", "progress", "note")
            )
            if (
                meta.get("stated", "") in annualized_stated
                and re.fullmatch(r"20\d{2}\s+year", horizon, flags=re.I)
                and (
                    not meta.get("actual", "")
                    or re.search(r"\brun[- ]rate\b|\bannualized program\b|matching annual horizon", blob, flags=re.I)
                )
            ):
                rows_to_delete.append(rr)

    for rr in sorted(set(rows_to_delete), reverse=True):
        ws.delete_rows(rr, 1)

def _ensure_cost_savings_run_rate_revision_row(ws: Any) -> None:
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    max_col = min(max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL), PROMISE_VISIBLE_MAX_COL)

    def _run_rate_actual_from_text(txt: Any) -> str:
        match = re.search(r"\$?\s*(\d+(?:\.\d+)?)\s*m\s+run[- ]rate", str(txt or ""), flags=re.I)
        return f"${float(match.group(1)):g}m run-rate" if match else ""

    run_rate_actual = ""
    target_text = ""
    for rr in range(1, int(ws.max_row or 0) + 1):
        if str(ws.cell(rr, 1).value or "").strip().lower() != "cost savings target":
            continue
        row_blob = " ".join(str(ws.cell(rr, cc).value or "") for cc in range(1, max_col + 1))
        run_rate_actual = run_rate_actual or _run_rate_actual_from_text(row_blob)
        if not target_text:
            for cc in (2, 3):
                candidate = str(ws.cell(rr, cc).value or "").strip()
                if re.search(r"\$\d", candidate):
                    target_text = candidate
                    break
    if not run_rate_actual:
        return

    latest_block_start = 0
    latest_block_end = 0
    cost_row_exists = False
    in_latest_block = False
    for rr in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        if _is_promise_section_row(ws, rr):
            if in_latest_block:
                break
            in_latest_block = first_txt == "2026-Q1 revisions"
            if in_latest_block:
                latest_block_start = rr
                latest_block_end = rr
            continue
        if not in_latest_block:
            continue
        if any(str(ws.cell(rr, cc).value or "").strip() for cc in range(1, max_col + 1)):
            latest_block_end = rr
        if first_txt.lower() == "cost savings target":
            cost_row_exists = True
    if not latest_block_start or cost_row_exists:
        return

    insert_at = latest_block_end + 1
    ws.insert_rows(insert_at, 1)
    template_row = latest_block_end if latest_block_end > latest_block_start else latest_block_start + 2
    for cc in range(1, max_col + 1):
        src = ws.cell(template_row, cc)
        dst = ws.cell(insert_at, cc)
        if src.has_style:
            dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
    row_values = [
        "Cost savings target",
        target_text or "$180m-$200m",
        target_text or "$180m-$200m",
        "Maintained",
        run_rate_actual,
        _promise_progress_label(run_rate_actual, metric="Cost savings target", stated="2026-Q1"),
        "On track",
        "Annualized program",
        "2026-Q1",
        "2026-03-31",
        "Latest disclosed run-rate against annualized savings target.",
    ]
    for cc, value in enumerate(row_values, start=1):
        ws.cell(insert_at, cc).value = value

def _remove_blank_promise_rows(ws: Any) -> None:
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    for merge_range in list(ws.merged_cells.ranges):
        if merge_range.min_row == merge_range.max_row:
            row_idx = int(merge_range.min_row)
            first_txt = str(ws.cell(row_idx, 1).value or "").strip()
            row_is_blank = all(
                str(ws.cell(row_idx, col_idx).value or "").strip() == ""
                for col_idx in range(1, PROMISE_VISIBLE_MAX_COL + 1)
            )
            if not (row_is_blank or first_txt in {"Metric", "Milestone", "Category"} or _is_promise_section_row(ws, row_idx)):
                continue
            try:
                ws.unmerge_cells(str(merge_range))
            except KeyError:
                pass
    rows_to_delete: List[int] = []
    inside_table = False
    current_section = ""
    suffixes = ("5B9BD5", "6FA8DC", "4472C4")
    for row_idx in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(row_idx, 1).value or "").strip()
        first_fill = str(ws.cell(row_idx, 1).fill.fgColor.rgb or "").upper()
        if first_txt and first_fill.endswith(suffixes):
            current_section = first_txt
            inside_table = False
            continue
        if first_txt in {"Metric", "Milestone", "Category"}:
            inside_table = True
            continue
        is_blank = all(str(ws.cell(row_idx, col_idx).value or "").strip() == "" for col_idx in range(1, PROMISE_VISIBLE_MAX_COL + 1))
        if is_blank and inside_table and current_section:
            if current_section.endswith("open guidance"):
                continue
            next_nonblank = None
            for look_ahead in range(row_idx + 1, int(ws.max_row or 0) + 1):
                if any(str(ws.cell(look_ahead, col_idx).value or "").strip() for col_idx in range(1, PROMISE_VISIBLE_MAX_COL + 1)):
                    next_nonblank = look_ahead
                    break
            if next_nonblank is not None and _is_promise_section_row(ws, next_nonblank):
                continue
            rows_to_delete.append(row_idx)
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx, 1)

def _repair_promise_table_header_merges(ws: Any) -> None:
    if ws is None or str(getattr(ws, "title", "")) != "Promise_Progress_UI":
        return
    max_col = PROMISE_VISIBLE_MAX_COL
    current_section = ""
    suffixes = ("5B9BD5", "6FA8DC", "4472C4")

    def _safe_unmerge(merge_range: Any) -> None:
        try:
            ws.unmerge_cells(str(merge_range))
        except (KeyError, ValueError):
            # Saved-workbook cleanup can encounter stale merged-cell metadata after
            # earlier insert/delete passes. Treat unmerge as idempotent.
            pass

    def _unmerge_row(row_idx: int) -> None:
        row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, max_col + 1)]
        for merge_range in list(ws.merged_cells.ranges):
            if merge_range.min_row <= row_idx <= merge_range.max_row:
                _safe_unmerge(merge_range)
        for col_idx, value in enumerate(row_values, start=1):
            if value not in (None, ""):
                ws.cell(row_idx, col_idx).value = value

    def _set_header_value(row_idx: int, col_idx: int, value: Any) -> None:
        for merge_range in list(ws.merged_cells.ranges):
            if (
                merge_range.min_row <= row_idx <= merge_range.max_row
                and merge_range.min_col <= col_idx <= merge_range.max_col
            ):
                _safe_unmerge(merge_range)
        try:
            ws.cell(row_idx, col_idx).value = value
        except AttributeError:
            if str(value or "").strip():
                raise

    def _style_header_row(row_idx: int) -> None:
        header_fill = copy(ws.cell(row_idx, 1).fill)
        if str(header_fill.fgColor.rgb or "").upper() in {"00000000", "00FFFFFF"}:
            header_fill = PatternFill("solid", fgColor="EAF3FB")
        header_font = copy(ws.cell(row_idx, 1).font)
        header_border = copy(ws.cell(row_idx, 1).border)
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = copy(header_fill)
            cell.font = copy(header_font)
            cell.border = copy(header_border)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for row_idx in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(row_idx, 1).value or "").strip()
        first_fill = str(ws.cell(row_idx, 1).fill.fgColor.rgb or "").upper()
        if first_txt and first_fill.endswith(suffixes):
            current_section = first_txt
            continue
        if first_txt == "Category":
            _unmerge_row(row_idx)
            labels = ["Category", "Score", "Evidence", "", "", "", "Read", "", "", "", "", ""]
            for col_idx, label in enumerate(labels, start=1):
                _set_header_value(row_idx, col_idx, label)
            _style_header_row(row_idx)
            ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=6)
            ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=max_col)
            continue
        if first_txt == "Milestone":
            _unmerge_row(row_idx)
            labels = ["Milestone", "Target / plan", "Actual", "Status", "Notes/source", "", "", "", "", "", "", ""]
            for col_idx, label in enumerate(labels, start=1):
                _set_header_value(row_idx, col_idx, label)
            _style_header_row(row_idx)
            ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=max_col)
            continue
        if first_txt != "Metric":
            if first_txt and current_section and current_section != "Management Credibility Scorecard":
                bad_full_row_merge = any(
                    merge_range.min_row == row_idx
                    and merge_range.max_row == row_idx
                    and merge_range.min_col == 1
                    and merge_range.max_col == max_col
                    for merge_range in ws.merged_cells.ranges
                )
                if bad_full_row_merge:
                    _unmerge_row(row_idx)
                if current_section.endswith("guidance progression"):
                    _unmerge_row(row_idx)
                    ws.merge_cells(start_row=row_idx, start_column=9, end_row=row_idx, end_column=max_col)
                elif (
                    current_section.endswith("open guidance")
                    or current_section == "Open guidance"
                    or current_section.endswith("milestone progression")
                ):
                    _unmerge_row(row_idx)
                    ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=max_col)
            continue
        if not current_section:
            _style_header_row(row_idx)
            continue
        _unmerge_row(row_idx)
        if current_section.endswith("guidance progression"):
            labels = ["Metric", "Initial guide", "Q1 update", "Q2 update", "Q3 update", "Q4 update", "Actual", "Status", "Notes/source", "", "", ""]
            merge_start = 9
        elif current_section.endswith("open guidance") or current_section == "Open guidance":
            labels = ["Metric", "Current guide", "Horizon", "Status", "Notes/source", "", "", "", "", "", "", ""]
            merge_start = 5
        elif current_section.endswith("milestone progression"):
            labels = ["Milestone", "Target / plan", "Actual", "Status", "Notes/source", "", "", "", "", "", "", ""]
            merge_start = 5
        else:
            labels = list(PROMISE_TIMELINE_HEADERS) + [""]
            merge_start = 0
        for col_idx, label in enumerate(labels, start=1):
            _set_header_value(row_idx, col_idx, label)
        _style_header_row(row_idx)
        if merge_start:
            ws.merge_cells(start_row=row_idx, start_column=merge_start, end_row=row_idx, end_column=max_col)

def _final_repair_promise_progress_ui(wb: Workbook, ticker: Any = "") -> None:
    """Last-mile Promise_Progress_UI repair after raw History_Q data exists."""
    if str(ticker or "").strip().upper() == "ANF":
        return
    if "Promise_Progress_UI" not in getattr(wb, "sheetnames", []):
        return
    ws = wb["Promise_Progress_UI"]
    if any(str(name).startswith("ANF_") for name in getattr(wb, "sheetnames", [])):
        return
    early_values = {
        str(ws.cell(rr, 1).value or "").strip()
        for rr in range(1, min(int(ws.max_row or 0), 120) + 1)
    }
    if {"Tariff impact", "Real estate activity"} & early_values:
        return
    if int(ws.max_row or 0) < 3:
        return

    def _norm_header(value: Any) -> str:
        txt = str(value or "").strip().lower()
        if txt in {"actual / latest actual", "actual / latest", "latest actual", "latest result"}:
            return "actual"
        return txt

    def _num(value: Any) -> Optional[float]:
        out = pd.to_numeric(value, errors="coerce")
        if pd.isna(out):
            return None
        val = float(out)
        return val if math.isfinite(val) else None

    def _history_labels(qd: date, row_map: Mapping[str, Any]) -> List[str]:
        labels = {str(qd)}
        fiscal_labels: Set[str] = set()
        fiscal_label = str(row_map.get("fiscal_label") or "").strip()
        if fiscal_label:
            fiscal_labels.add(fiscal_label)
        fy = _num(row_map.get("fiscal_year"))
        fq = _num(row_map.get("fiscal_quarter"))
        if fy is not None and fq is not None and 1 <= int(fq) <= 4:
            fiscal_labels.add(f"{int(fy)}-Q{int(fq)}")
        if fiscal_labels:
            labels.update(fiscal_labels)
        else:
            labels.add(f"{qd.year}-Q{((qd.month - 1) // 3) + 1}")
        return [x for x in labels if x]

    actuals_by_period: Dict[str, Dict[str, float]] = {}
    actuals_by_year: Dict[int, Dict[str, float]] = {}
    year_end_dates: Dict[int, date] = {}
    history_period_labels_by_date: Dict[date, Set[str]] = {}

    def _add_period(labels: Iterable[str], key: str, value: Optional[float]) -> None:
        if value is None:
            return
        for label in labels:
            actuals_by_period.setdefault(label, {})[key] = value

    def _add_year(year: int, key: str, value: Optional[float]) -> None:
        if value is None:
            return
        actuals_by_year.setdefault(int(year), {})[key] = actuals_by_year.setdefault(int(year), {}).get(key, 0.0) + value

    def _labels_year_or_calendar(qd: date, fiscal_year: Optional[float], fiscal_quarter: Optional[float]) -> Tuple[Set[str], int]:
        labels: Set[str] = {str(qd)}
        if fiscal_year is not None and fiscal_quarter is not None and 1 <= int(fiscal_quarter) <= 4:
            labels.add(f"{int(fiscal_year)}-Q{int(fiscal_quarter)}")
            return labels, int(fiscal_year)
        hist_labels = history_period_labels_by_date.get(qd)
        if hist_labels:
            labels.update(hist_labels)
            year_match = next((re.match(r"^(20\d{2})-Q[1-4]$", label) for label in sorted(hist_labels)), None)
            if year_match:
                return labels, int(year_match.group(1))
        labels.add(f"{qd.year}-Q{((qd.month - 1) // 3) + 1}")
        return labels, int(qd.year)

    if "History_Q" in getattr(wb, "sheetnames", []):
        hist_ws = wb["History_Q"]
        headers = {
            re.sub(r"[^a-z0-9]+", "_", str(hist_ws.cell(1, cc).value or "").strip().lower()).strip("_"): cc
            for cc in range(1, int(hist_ws.max_column or 0) + 1)
        }
        q_col = headers.get("quarter")
        if q_col:
            for rr in range(2, int(hist_ws.max_row or 0) + 1):
                qd = _date_or_none(hist_ws.cell(rr, q_col).value)
                if qd is None:
                    continue
                row_map = {name: hist_ws.cell(rr, col_idx).value for name, col_idx in headers.items()}
                labels = _history_labels(qd, row_map)
                history_period_labels_by_date.setdefault(qd, set()).update(
                    label for label in labels if re.fullmatch(r"20\d{2}-Q[1-4]", str(label))
                )
                fiscal_year = _num(row_map.get("fiscal_year"))
                year = int(fiscal_year) if fiscal_year is not None else int(qd.year)
                year_end_dates[year] = max(year_end_dates.get(year, date.min), qd)
                revenue = _num(row_map.get("revenue"))
                op_income = _num(row_map.get("op_income"))
                cfo = _num(row_map.get("cfo"))
                capex = _num(row_map.get("capex"))
                eps = _num(row_map.get("eps_diluted"))
                shares = _num(row_map.get("shares_diluted"))
                buybacks = _num(row_map.get("buybacks_cash"))
                operating_margin = (op_income / revenue) if revenue and op_income is not None else None
                fcf = (cfo - capex) if cfo is not None and capex is not None else None
                for key, value in (
                    ("revenue", revenue),
                    ("operating_margin", operating_margin),
                    ("fcf", fcf),
                    ("capex", capex),
                    ("eps", eps),
                    ("shares", shares),
                    ("buybacks", buybacks),
                ):
                    _add_period(labels, key, value)
                    _add_year(year, key, value)

    if "Adjusted_Metrics" in getattr(wb, "sheetnames", []):
        adj_ws = wb["Adjusted_Metrics"]
        headers = {
            re.sub(r"[^a-z0-9]+", "_", str(adj_ws.cell(1, cc).value or "").strip().lower()).strip("_"): cc
            for cc in range(1, int(adj_ws.max_column or 0) + 1)
        }
        q_col = headers.get("quarter")
        if q_col:
            for rr in range(2, int(adj_ws.max_row or 0) + 1):
                qd = _date_or_none(adj_ws.cell(rr, q_col).value)
                if qd is None:
                    continue
                period_type = str(adj_ws.cell(rr, headers.get("period_type", 0)).value or "").strip().lower() if headers.get("period_type") else ""
                if period_type and "annual" in period_type:
                    continue
                row_map = {name: adj_ws.cell(rr, col_idx).value for name, col_idx in headers.items()}
                fiscal_year = _num(row_map.get("fiscal_year"))
                fiscal_quarter = _num(row_map.get("fiscal_quarter"))
                labels, year = _labels_year_or_calendar(qd, fiscal_year, fiscal_quarter)
                year_end_dates[year] = max(year_end_dates.get(year, date.min), qd)
                for key, source_name in (
                    ("adj_ebit", "adj_ebit"),
                    ("adj_ebitda", "adj_ebitda"),
                    ("adj_eps", "adj_eps"),
                    ("adj_fcf", "adj_fcf"),
                ):
                    value = _num(row_map.get(source_name))
                    _add_period(labels, key, value)
                    _add_year(year, key, value)

    if "Valuation" in getattr(wb, "sheetnames", []):
        val_ws = wb["Valuation"]
        qcols: Dict[str, int] = {}
        for cc in range(1, int(val_ws.max_column or 0) + 1):
            label = str(val_ws.cell(6, cc).value or "").strip()
            if re.fullmatch(r"20\d{2}-Q[1-4]", label):
                qcols[label] = cc
        val_row_by_label = {
            str(val_ws.cell(rr, 1).value or "").strip(): rr
            for rr in range(1, int(val_ws.max_row or 0) + 1)
            if str(val_ws.cell(rr, 1).value or "").strip()
        }
        for row_label, actual_key, scale in (
            ("Adj EPS", "adj_eps", 1.0),
            ("Adj EBITDA", "adj_ebitda", 1_000_000.0),
        ):
            row_idx = val_row_by_label.get(row_label)
            if not row_idx:
                continue
            for label, col_idx in qcols.items():
                vv = _num(val_ws.cell(row_idx, col_idx).value)
                if vv is None:
                    continue
                year_match = re.match(r"^(20\d{2})-Q([1-4])$", label)
                if not year_match:
                    continue
                value = float(vv) * float(scale)
                _add_period([label], actual_key, value)
        for row_label, actual_key, scale in (
            ("Adj EBIT (TTM)", "adj_ebit", 1_000_000.0),
            ("Adj EBITDA (TTM)", "adj_ebitda", 1_000_000.0),
            ("Adj EPS (TTM)", "adj_eps", 1.0),
        ):
            row_idx = val_row_by_label.get(row_label)
            if not row_idx:
                continue
            for label, col_idx in qcols.items():
                year_match = re.match(r"^(20\d{2})-Q4$", label)
                if not year_match:
                    continue
                vv = _num(val_ws.cell(row_idx, col_idx).value)
                if vv is None:
                    continue
                actuals_by_year.setdefault(int(year_match.group(1)), {})[actual_key] = float(vv) * float(scale)

    def _metric_actual_key(metric_in: Any) -> str:
        metric_low = str(metric_in or "").strip().lower()
        if "adjusted ebitda" in metric_low or "adj ebitda" in metric_low:
            return "adj_ebitda"
        if "adjusted ebit" in metric_low or "adj ebit" in metric_low:
            return "adj_ebit"
        if "fcf" in metric_low or "free cash" in metric_low:
            return "adj_fcf" if "adjusted" in metric_low or "adj" in metric_low else "fcf"
        if "capex" in metric_low or "capital expenditure" in metric_low:
            return "capex"
        if "operating margin" in metric_low or "op margin" in metric_low:
            return "operating_margin"
        if "eps" in metric_low:
            return "adj_eps" if "adjusted" in metric_low or "adj" in metric_low else "eps"
        if "share" in metric_low and "repurchase" not in metric_low and "buyback" not in metric_low:
            return "shares"
        if "repurchase" in metric_low or "buyback" in metric_low:
            return "buybacks"
        if "revenue" in metric_low or "sales" in metric_low:
            return "revenue"
        return ""

    def _format_actual(metric_in: Any, key: str, value: Any) -> str:
        val = pd.to_numeric(value, errors="coerce")
        if pd.isna(val):
            return ""
        num = float(val)
        if key == "operating_margin":
            return f"{num * 100:.1f}%"
        if key in {"adj_eps", "eps"}:
            return f"${num:,.2f}"
        if key == "shares":
            return f"{num / 1_000_000:,.1f}m" if abs(num) > 1_000_000 else f"{num:,.1f}m"
        abs_num = abs(num)
        if abs_num >= 1_000_000_000:
            return f"${num / 1_000_000_000:,.2f}bn"
        if abs_num >= 1_000_000:
            return f"${num / 1_000_000:,.1f}m"
        return f"${num:,.1f}"

    def _quarter_score(label: Any) -> Tuple[Optional[int], Optional[int]]:
        m = re.search(r"\b(20\d{2})-Q([1-4])\b", str(label or ""), flags=re.I)
        if not m:
            return None, None
        return int(m.group(1)), int(m.group(2))

    def _annual_year(label: Any) -> Optional[int]:
        m = re.fullmatch(r"(20\d{2})\s+year", str(label or "").strip(), flags=re.I)
        return int(m.group(1)) if m else None

    def _actual_for_period(metric: Any, period_label: Any) -> str:
        key = _metric_actual_key(metric)
        if not key:
            return ""
        vals = actuals_by_period.get(str(period_label or "").strip(), {})
        if key in vals:
            return _format_actual(metric, key, vals[key])
        if key == "eps" and "adj_eps" in vals:
            return _format_actual(metric, "adj_eps", vals["adj_eps"])
        if key == "adj_fcf" and "fcf" in vals:
            return _format_actual(metric, "fcf", vals["fcf"])
        return ""

    def _actual_for_year(metric: Any, year: int) -> str:
        key = _metric_actual_key(metric)
        if not key:
            return ""
        vals = actuals_by_year.get(int(year), {})
        if key in vals:
            return _format_actual(metric, key, vals[key])
        if key == "eps" and "adj_eps" in vals:
            return _format_actual(metric, "adj_eps", vals["adj_eps"])
        if key == "adj_fcf" and "fcf" in vals:
            return _format_actual(metric, "fcf", vals["fcf"])
        return ""

    def _ytd_for_year_to_quarter(metric: Any, year: int, quarter_num: int) -> str:
        key = _metric_actual_key(metric)
        if not key or quarter_num not in {1, 2, 3, 4}:
            return ""
        if key in {"operating_margin", "shares"}:
            return ""
        vals: List[float] = []
        actual_key = key
        for idx in range(1, int(quarter_num) + 1):
            period_vals = actuals_by_period.get(f"{int(year)}-Q{idx}", {})
            if actual_key not in period_vals and actual_key == "eps" and "adj_eps" in period_vals:
                actual_key = "adj_eps"
            if actual_key not in period_vals and actual_key == "adj_fcf" and "fcf" in period_vals:
                actual_key = "fcf"
            if actual_key not in period_vals:
                return ""
            vv = _num(period_vals.get(actual_key))
            if vv is None:
                return ""
            vals.append(float(vv))
        return _format_actual(metric, actual_key, sum(vals))

    def _numbers(txt: Any) -> List[float]:
        out: List[float] = []
        for match in re.findall(r"(?<![A-Za-z])\d+(?:,\d{3})*(?:\.\d+)?", str(txt or "")):
            try:
                out.append(float(match.replace(",", "")))
            except Exception:
                continue
        return out

    def _status_from_guidance_actual(metric: Any, guide: Any, actual: Any) -> str:
        guide_nums = _numbers(guide)
        actual_nums = _numbers(actual)
        if not actual_nums:
            return ""
        if not guide_nums:
            return "Completed"
        metric_low = str(metric or "").lower()
        actual_val = actual_nums[0]
        lo = min(guide_nums)
        hi = max(guide_nums)
        if len(guide_nums) == 1:
            lo = hi = guide_nums[0]
        tol = max(abs(hi) * 0.001, 0.001)
        if "capex" in metric_low or ("share" in metric_low and "repurchase" not in metric_low and "buyback" not in metric_low):
            return "Hit" if lo - tol <= actual_val <= hi + tol else "Mixed"
        if actual_val > hi + tol:
            return "Beat"
        if actual_val < lo - tol:
            return "Missed"
        return "Hit"

    def _run_rate_actual_from_text(txt: Any) -> str:
        match = re.search(r"\$?\s*(\d+(?:\.\d+)?)\s*m\s+run[- ]rate", str(txt or ""), flags=re.I)
        return f"${float(match.group(1)):g}m run-rate" if match else ""

    max_col = int(ws.max_column or 0)
    current_block = ""
    active_cols: Dict[str, int] = {}
    rows_to_delete: List[int] = []
    annual_latest_by_metric: Dict[Tuple[int, str], Dict[str, Any]] = {}
    annual_q4_rows: Set[Tuple[int, str]] = set()
    q4_blocks: Dict[int, Dict[str, Any]] = {}

    def _header_map(row_idx: int) -> Dict[str, int]:
        return {
            _norm_header(ws.cell(row_idx, cc).value): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(row_idx, cc).value or "").strip()
        }

    for rr in range(1, int(ws.max_row or 0) + 1):
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        first_fill = str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper()
        row_map = _header_map(rr)
        if "actual" in row_map and ("metric" in row_map or "milestone" in row_map):
            active_cols = row_map
            continue
        if first_txt and (first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")) or first_txt.endswith("revisions")):
            current_block = first_txt
            active_cols = {}
            block_year, block_q = _quarter_score(current_block)
            if block_year and block_q == 4:
                q4_blocks.setdefault(
                    block_year,
                    {
                        "start": rr,
                        "end": rr,
                        "source_date": str(year_end_dates.get(block_year) or date(block_year, 12, 31)),
                    },
                )
            continue
        if not active_cols:
            continue
        metric_col = active_cols.get("metric") or active_cols.get("milestone")
        if not metric_col:
            continue
        metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
        if not metric_txt or metric_txt.lower() in {"metric", "milestone"}:
            continue
        useful_values = [
            ws.cell(rr, cc).value
            for cc in range(1, min(max_col, PROMISE_VISIBLE_MAX_COL) + 1)
            if cc != metric_col
        ]
        if all(str(value or "").strip() == "" for value in useful_values):
            rows_to_delete.append(rr)
            continue

        actual_col = active_cols.get("actual")
        progress_col = active_cols.get("progress / run-rate")
        status_col = active_cols.get("status")
        note_col = active_cols.get("source / note") or active_cols.get("notes/source")
        actual_txt = str(ws.cell(rr, actual_col).value or "").strip() if actual_col else ""
        note_txt = str(ws.cell(rr, note_col).value or "").strip() if note_col else ""
        if metric_txt.lower() == "debt reduction" and actual_col and status_col:
            status_txt = str(ws.cell(rr, status_col).value or "").strip().lower()
            if not actual_txt and status_txt in {"completed", "hit"} and re.search(r"\b(repay|repaid|paid off|pay down|paydown)\b", note_txt, flags=re.I):
                ws.cell(rr, actual_col).value = "Debt repaid"
            continue
        if metric_txt.lower() == "cost savings target" and actual_col:
            run_rate = _run_rate_actual_from_text(" ".join([actual_txt, note_txt]))
            if run_rate:
                ws.cell(rr, actual_col).value = ""
                if progress_col:
                    ws.cell(rr, progress_col).value = _promise_progress_label(run_rate, metric=metric_txt, stated=current_block)
            continue

        horizon_col = active_cols.get("horizon")
        stated_col = active_cols.get("stated in")
        new_col = active_cols.get("new/current guide")
        prev_col = active_cols.get("previous guide")
        source_col = active_cols.get("source date")
        if not (horizon_col and stated_col and new_col and status_col):
            continue
        horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip()
        stated_txt = str(ws.cell(rr, stated_col).value or current_block).strip()
        source_txt = str(ws.cell(rr, source_col).value or "").strip() if source_col else ""
        year = _annual_year(horizon_txt)
        stated_year, stated_q = _quarter_score(stated_txt)
        if year is None or stated_year is None:
            continue
        if stated_year == year and stated_q in {1, 2, 3} and actual_col:
            quarter_actual = _actual_for_period(metric_txt, stated_txt)
            if quarter_actual:
                ws.cell(rr, actual_col).value = quarter_actual
                if progress_col:
                    ytd_actual = _ytd_for_year_to_quarter(metric_txt, year, stated_q)
                    ws.cell(rr, progress_col).value = f"YTD: {ytd_actual}" if ytd_actual else _promise_progress_label(quarter_actual, metric=metric_txt, stated=stated_txt)
                ws.cell(rr, status_col).value = "On track"
            elif actual_txt and re.search(r"\bbn|\$|m\b|\d", actual_txt, flags=re.I):
                ws.cell(rr, actual_col).value = ""
                ws.cell(rr, status_col).value = "Open"
        elif stated_year == year and stated_q == 4:
            annual_actual = _actual_for_year(metric_txt, year)
            actual_key_for_row = _metric_actual_key(metric_txt)
            force_annual_actual = (
                actual_key_for_row in {"revenue", "adj_ebit", "adj_ebitda", "fcf", "capex", "buybacks"}
                and not re.search(r"\b(growth|margin|rate|bps|basis|share count|diluted shares)\b", metric_txt, flags=re.I)
            )
            if annual_actual and actual_col and (not actual_txt or force_annual_actual):
                ws.cell(rr, actual_col).value = annual_actual
                guide = str(ws.cell(rr, new_col).value or (ws.cell(rr, prev_col).value if prev_col else "") or "").strip()
                status = _status_from_guidance_actual(metric_txt, guide, annual_actual)
                if status:
                    ws.cell(rr, status_col).value = status
            annual_q4_rows.add((year, metric_txt.lower()))
        elif stated_year < year and actual_col and actual_txt and re.search(r"\bbn|\$|m\b|\d", actual_txt, flags=re.I):
            ws.cell(rr, actual_col).value = ""
            ws.cell(rr, status_col).value = "Open"

        if stated_year <= year and (stated_year < year or stated_q in {1, 2, 3}):
            guide = str(ws.cell(rr, new_col).value or (ws.cell(rr, prev_col).value if prev_col else "") or "").strip()
            if guide and not re.search(r"\bactual\b", guide, flags=re.I):
                key = (year, metric_txt.lower())
                try:
                    source_dt = pd.Timestamp(source_txt).date()
                except Exception:
                    source_dt = date.min
                prior = annual_latest_by_metric.get(key)
                if prior is None or source_dt >= prior.get("source_dt", date.min):
                    annual_latest_by_metric[key] = {"metric": metric_txt, "guide": guide, "source_dt": source_dt}

        if str(current_block).endswith("revisions"):
            block_year, block_q = _quarter_score(current_block)
            if block_year and block_q == 4:
                q4_blocks[block_year] = {
                    "start": q4_blocks.get(block_year, {}).get("start", rr),
                    "end": rr,
                    "source_date": source_txt or str(year_end_dates.get(block_year) or date(block_year, 12, 31)),
                }

    for rr in sorted(set(rows_to_delete), reverse=True):
        ws.delete_rows(rr, 1)

    for merge_range in list(ws.merged_cells.ranges):
        if merge_range.min_row != merge_range.max_row:
            continue
        rr = int(merge_range.min_row)
        first_val = str(ws.cell(rr, 1).value or "").strip()
        first_fill = str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper()
        if (
            first_val
            and first_val not in {"Metric", "Milestone", "Category"}
            and not first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4"))
        ):
            ws.unmerge_cells(str(merge_range))

    if str(ticker or "").strip().upper() == "PBI":
        run_rate_actual = ""
        target_text = ""
        for rr in range(1, int(ws.max_row or 0) + 1):
            if str(ws.cell(rr, 1).value or "").strip().lower() != "cost savings target":
                continue
            row_blob = " ".join(str(ws.cell(rr, cc).value or "") for cc in range(1, min(max_col, 10) + 1))
            run_rate_actual = run_rate_actual or _run_rate_actual_from_text(row_blob)
            target_text = target_text or str(ws.cell(rr, 2).value or "").strip()
        if run_rate_actual:
            latest_block_start = 0
            latest_block_end = 0
            in_latest_block = False
            cost_row_exists = False
            for rr in range(1, int(ws.max_row or 0) + 1):
                first_txt = str(ws.cell(rr, 1).value or "").strip()
                first_fill = str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper()
                is_section = bool(first_txt) and (first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")) or first_txt.endswith("revisions"))
                if is_section:
                    if in_latest_block:
                        break
                    in_latest_block = first_txt == "2026-Q1 revisions"
                    if in_latest_block:
                        latest_block_start = rr
                        latest_block_end = rr
                    continue
                if not in_latest_block:
                    continue
                if any(str(ws.cell(rr, cc).value or "").strip() for cc in range(1, min(max_col, 10) + 1)):
                    latest_block_end = rr
                if first_txt.lower() == "cost savings target":
                    cost_row_exists = True
            if latest_block_start and not cost_row_exists:
                insert_at = latest_block_end + 1
                ws.insert_rows(insert_at, 1)
                template_row = latest_block_end if latest_block_end > latest_block_start else latest_block_start + 2
                for merge_range in list(ws.merged_cells.ranges):
                    if merge_range.min_row == insert_at and merge_range.max_row == insert_at:
                        ws.unmerge_cells(str(merge_range))
                for cc in range(1, min(max_col, 10) + 1):
                    src = ws.cell(template_row, cc)
                    dst = ws.cell(insert_at, cc)
                    if src.has_style:
                        dst._style = copy(src._style)
                    dst.font = copy(src.font)
                    dst.fill = copy(src.fill)
                    dst.border = copy(src.border)
                    dst.alignment = copy(src.alignment)
                    dst.number_format = src.number_format
                row_values = [
                    "Cost savings target",
                    target_text or "$180m-$200m",
                    target_text or "$180m-$200m",
                    "Maintained",
                    run_rate_actual,
                    _promise_progress_label(run_rate_actual, metric="Cost savings target", stated="2026-Q1"),
                    "On track",
                    "Annualized program",
                    "2026-Q1",
                    "2026-03-31",
                    "Latest disclosed run-rate against annualized savings target.",
                ]
                for cc, value in enumerate(row_values, start=1):
                    ws.cell(insert_at, cc).value = value

    # Final guard after all insertions: remove any metric-label stubs whose
    # data cells are empty, and clear same-row merges left behind by deletions.
    cleanup_rows: List[int] = []
    active_cols = {}
    for rr in range(1, int(ws.max_row or 0) + 1):
        row_map = _header_map(rr)
        first_txt = str(ws.cell(rr, 1).value or "").strip()
        first_fill = str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper()
        if first_txt and (first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")) or first_txt.endswith("revisions")):
            active_cols = {}
            continue
        if "actual" in row_map and ("metric" in row_map or "milestone" in row_map):
            active_cols = row_map
            continue
        metric_col = active_cols.get("metric") or active_cols.get("milestone") if active_cols else None
        if not metric_col:
            continue
        metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
        if not metric_txt or metric_txt.lower() in {"metric", "milestone"}:
            continue
        if all(str(ws.cell(rr, cc).value or "").strip() == "" for cc in range(1, min(max_col, 10) + 1) if cc != metric_col):
            cleanup_rows.append(rr)
    for rr in sorted(set(cleanup_rows), reverse=True):
        ws.delete_rows(rr, 1)
    for merge_range in list(ws.merged_cells.ranges):
        if merge_range.min_row == merge_range.max_row:
            row_has_text = any(
                str(ws.cell(int(merge_range.min_row), cc).value or "").strip()
                for cc in range(1, min(max_col, 10) + 1)
            )
            if not row_has_text:
                ws.unmerge_cells(str(merge_range))
    _remove_empty_promise_revision_blocks(ws)
    _polish_promise_scorecard_layout(ws)
    _apply_source_backed_promise_mapping_overrides(wb, ticker)
    _finalize_promise_revision_semantics(ws)
    _apply_promise_grid_style(ws)

def insert_management_credibility_scorecard(
    deps: PromiseProgressWorksheetRepairDeps,
    ws: Any,
    ticker: Any = "",
) -> None:
    _set_runtime(deps)
    return _insert_management_credibility_scorecard(ws, ticker)


def polish_promise_scorecard_layout(
    deps: PromiseProgressWorksheetRepairDeps,
    ws: Any,
) -> None:
    _set_runtime(deps)
    return _polish_promise_scorecard_layout(ws)


def repair_promise_table_header_merges(
    deps: PromiseProgressWorksheetRepairDeps,
    ws: Any,
) -> None:
    _set_runtime(deps)
    return _repair_promise_table_header_merges(ws)


def final_repair_promise_progress_ui(
    deps: PromiseProgressWorksheetRepairDeps,
    wb: Any,
    ticker: Any = "",
) -> None:
    _set_runtime(deps)
    return _final_repair_promise_progress_ui(wb, ticker)

