"""Promise source-backed override worksheet mutator."""
from __future__ import annotations

import math
import re
from copy import copy
from dataclasses import dataclass
from typing import Any, Dict, List, Mapping, MutableMapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .excel_writer_promise_source_overrides import (
    PromiseSourceOverrideSupport,
    PromiseSourceOverrideSupportDeps,
)


@dataclass(frozen=True)
class PromiseSourceOverrideMutatorDeps:
    runtime: MutableMapping[str, Any]


def apply_source_backed_promise_mapping_overrides(
    deps: PromiseSourceOverrideMutatorDeps,
) -> None:
    """Apply curated source-backed Promise fixes after generic timeline rewriting.

    The generic promise builder intentionally avoids raw transcript fragments, but
    a few company-defined metrics need exact source semantics after rows are
    normalized.  Keep this function small and auditable: it only changes rows
    where the underlying source definition is known and visible.
    """
    runtime = deps.runtime
    wb = runtime["wb"]
    ticker = runtime.get("ticker", "")
    _date_or_none = runtime["_date_or_none"]
    _set_promise_row_semantics = runtime["_set_promise_row_semantics"]
    PROMISE_TIMELINE_HEADERS = runtime["PROMISE_TIMELINE_HEADERS"]
    _ensure_anf_promise_hidden_source_keys = runtime["_ensure_anf_promise_hidden_source_keys"]
    _promise_stated_quarter_parts = runtime["_promise_stated_quarter_parts"]
    _promise_annual_year = runtime["_promise_annual_year"]
    _promise_progress_label = runtime["_promise_progress_label"]
    _promise_value_looks_like_progress = runtime["_promise_value_looks_like_progress"]
    _remove_pbi_duplicate_cost_savings_timeline_rows = runtime["_remove_pbi_duplicate_cost_savings_timeline_rows"]
    ticker_txt = str(ticker or "").strip().upper()
    if "Promise_Progress_UI" not in getattr(wb, "sheetnames", []):
        return
    ws = wb["Promise_Progress_UI"]
    max_col = max(10, int(ws.max_column or 0))
    promise_source_support = PromiseSourceOverrideSupport(
        PromiseSourceOverrideSupportDeps(runtime={**globals(), **locals()})
    )
    _promise_override_lifecycle_id = promise_source_support.lifecycle_id

    def _norm_header(value: Any) -> str:
        txt = str(value or "").strip().lower()
        if txt in {"actual / latest actual", "actual / latest", "latest actual", "latest result"}:
            return "actual"
        return txt

    def _sectionish(row_idx: int) -> bool:
        first_txt = str(ws.cell(row_idx, 1).value or "").strip()
        if not first_txt:
            return False
        first_fill = str(ws.cell(row_idx, 1).fill.fgColor.rgb or "").upper()
        return (
            first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4"))
            or first_txt.endswith("revisions")
            or first_txt.endswith("guidance progression")
            or first_txt.endswith("open guidance")
            or first_txt.endswith("milestone progression")
            or first_txt.endswith("timeline / revision log")
        )

    def _header_map(row_idx: int) -> Dict[str, int]:
        return {
            _norm_header(ws.cell(row_idx, cc).value): cc
            for cc in range(1, max_col + 1)
            if str(ws.cell(row_idx, cc).value or "").strip()
        }

    def _promise_rows() -> List[Tuple[int, str, Dict[str, int]]]:
        out: List[Tuple[int, str, Dict[str, int]]] = []
        current_section = ""
        active_cols: Dict[str, int] = {}
        for rr in range(1, int(ws.max_row or 0) + 1):
            first_txt = str(ws.cell(rr, 1).value or "").strip()
            if _sectionish(rr):
                current_section = first_txt
                active_cols = {}
                continue
            row_map = _header_map(rr)
            if "metric" in row_map or "milestone" in row_map:
                active_cols = row_map
                continue
            metric_col = active_cols.get("metric") or active_cols.get("milestone")
            if metric_col:
                metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
                if metric_txt and metric_txt.lower() not in {"metric", "milestone"}:
                    out.append((rr, current_section, dict(active_cols)))
        return out

    def _copy_row_style(src_row: int, dst_row: int) -> None:
        for cc in range(1, max_col + 1):
            src = ws.cell(src_row, cc)
            dst = ws.cell(dst_row, cc)
            if src.has_style:
                dst._style = copy(src._style)
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height or 22.0

    def _block_bounds(section_name: str) -> Tuple[int, int, int]:
        start = 0
        for rr in range(1, int(ws.max_row or 0) + 1):
            if str(ws.cell(rr, 1).value or "").strip() == section_name and _sectionish(rr):
                start = rr
                break
        if not start:
            return 0, 0, 0
        header = 0
        end = int(ws.max_row or 0)
        for rr in range(start + 1, int(ws.max_row or 0) + 1):
            first_txt = str(ws.cell(rr, 1).value or "").strip()
            if _sectionish(rr):
                end = rr - 1
                break
            row_map = _header_map(rr)
            if not header and {"metric", "previous guide", "new/current guide"}.issubset(set(row_map)):
                header = rr
        return start, header, end

    def _revision_score(section_name: str) -> int:
        m = re.search(r"\b(20\d{2})-Q([1-4])\b", str(section_name or ""), flags=re.I)
        return int(m.group(1)) * 10 + int(m.group(2)) if m else 0

    _promise_display_section_from_horizon = promise_source_support.display_section_from_horizon
    _source_date_ordinal = promise_source_support.source_date_ordinal
    _append_prior_source_note = promise_source_support.append_prior_source_note

    def _ensure_timeline_block(section_name: str) -> None:
        start, header, _ = _block_bounds(section_name)
        if start and header:
            return
        target_score = _revision_score(section_name)
        insert_at = int(ws.max_row or 0) + 1
        for rr in range(1, int(ws.max_row or 0) + 1):
            title = str(ws.cell(rr, 1).value or "").strip()
            if not title.endswith("revisions") or not _sectionish(rr):
                continue
            if _revision_score(title) < target_score:
                insert_at = rr
                break
        ws.insert_rows(insert_at, 2)
        ws.cell(insert_at, 1).value = section_name
        ws.cell(insert_at, 1).fill = PatternFill("solid", fgColor="5B9BD5")
        for cc, value in enumerate(PROMISE_TIMELINE_HEADERS, start=1):
            ws.cell(insert_at + 1, cc).value = value

    def _upsert_timeline_row(section_name: str, values: Sequence[Any]) -> None:
        values = list(values)
        if len(values) == 10:
            values = values[:5] + [""] + values[5:]
        if len(values) >= 9:
            section_name = _promise_display_section_from_horizon(values[8], values[7] if len(values) > 7 else "", section_name)
        if len(values) >= 11 and _promise_value_looks_like_progress(values[4], metric=values[0]):
            values[5] = values[5] or _promise_progress_label(values[4], metric=values[0], stated=values[8] if len(values) > 8 else "")
            if str(values[0] or "").strip() != "Cost savings target":
                values[4] = ""
                if str(values[6] or "").strip().lower() in {"completed", "hit", "missed", "beat"}:
                    values[6] = "On track"
        _ensure_timeline_block(section_name)
        start, header, end = _block_bounds(section_name)
        if not start or not header:
            return
        target_row = 0
        merge_existing: Dict[str, Any] = {}
        existing_note_to_preserve = ""
        incoming_metric = str(values[0] if len(values) > 0 else "").strip()
        incoming_horizon = str(values[7] if len(values) > 7 else "").strip()
        incoming_stated = str(values[8] if len(values) > 8 else "").strip()
        incoming_source_ord = _source_date_ordinal(values[9] if len(values) > 9 else "")
        for rr, section, cols in _promise_rows():
            if section != section_name:
                continue
            metric_txt = str(ws.cell(rr, cols.get("metric", 1)).value or "").strip()
            stated_txt = str(ws.cell(rr, cols.get("stated in", 0)).value or "").strip() if cols.get("stated in") else ""
            if metric_txt == incoming_metric and stated_txt == incoming_stated:
                target_row = rr
                note_col = cols.get("source / note") or cols.get("notes/source")
                if note_col:
                    existing_note_to_preserve = str(ws.cell(rr, note_col).value or "").strip()
                break
        if not target_row:
            for rr, section, cols in _promise_rows():
                if section != section_name:
                    continue
                metric_txt = str(ws.cell(rr, cols.get("metric", 1)).value or "").strip()
                horizon_txt = str(ws.cell(rr, cols.get("horizon", 0)).value or "").strip() if cols.get("horizon") else ""
                if metric_txt != incoming_metric or horizon_txt != incoming_horizon:
                    continue
                target_row = rr
                merge_existing = {
                    name: ws.cell(rr, col).value
                    for name, col in cols.items()
                    if name in {"previous guide", "new/current guide", "stated in", "source date", "source / note"}
                }
                existing_source_ord = _source_date_ordinal(merge_existing.get("source date"))
                if existing_source_ord >= incoming_source_ord and incoming_source_ord:
                    prev_col = cols.get("previous guide")
                    if prev_col and not str(ws.cell(rr, prev_col).value or "").strip() and len(values) > 2:
                        ws.cell(rr, prev_col).value = values[2]
                    note_col = cols.get("source / note") or cols.get("notes/source")
                    if note_col:
                        existing_note = str(ws.cell(rr, note_col).value or "").strip()
                        tmp_values = [""] * 11
                        tmp_values[10] = existing_note
                        _append_prior_source_note(tmp_values, {
                            "new/current guide": values[2] if len(values) > 2 else "",
                            "stated in": incoming_stated,
                            "source date": values[9] if len(values) > 9 else "",
                        })
                        ws.cell(rr, note_col).value = tmp_values[10]
                    return
                break
        if not target_row:
            insert_at = end + 1
            ws.insert_rows(insert_at, 1)
            template = max(header + 1, end)
            if template >= insert_at:
                template = header
            _copy_row_style(template, insert_at)
            target_row = insert_at
        for merge_range in list(ws.merged_cells.ranges):
            if merge_range.min_row == target_row and merge_range.max_row == target_row:
                ws.unmerge_cells(str(merge_range))
        for cc, value in enumerate(values, start=1):
            ws.cell(target_row, cc).value = value
        if merge_existing:
            _append_prior_source_note(values, merge_existing)
            if len(values) >= 11:
                ws.cell(target_row, 11).value = values[10]
        elif existing_note_to_preserve and "initial guide" in existing_note_to_preserve.lower() and len(values) >= 11:
            current_note = str(values[10] or "").strip()
            preserved_match = re.search(r"\bInitial guide\b[^.]*\.", existing_note_to_preserve, flags=re.I)
            preserved_note = preserved_match.group(0).strip() if preserved_match else existing_note_to_preserve
            if preserved_note and preserved_note.lower() not in current_note.lower():
                ws.cell(target_row, 11).value = f"{preserved_note} {current_note}".strip()
        hidden_id = _promise_override_lifecycle_id(values[0] if len(values) > 0 else "", values[7] if len(values) > 7 else "")
        ws.cell(target_row, 15).value = hidden_id
        ws.column_dimensions["O"].hidden = True

    def _upsert_guidance_progression_section(year: int, rows: Sequence[Mapping[str, Any]]) -> None:
        if not rows:
            return
        title = f"{int(year)} guidance progression"
        start, header, end = _block_bounds(title)
        insert_at = start or 0
        if start:
            for merge_range in list(ws.merged_cells.ranges):
                if merge_range.min_row <= end and merge_range.max_row >= start:
                    ws.unmerge_cells(str(merge_range))
            ws.delete_rows(start, max(end - start + 1, 1))
        else:
            target_year = int(year)
            insert_at = int(ws.max_row or 0) + 1
            fallback_insert: Optional[int] = None
            for rr in range(1, int(ws.max_row or 0) + 1):
                first_txt = str(ws.cell(rr, 1).value or "").strip()
                if first_txt.endswith("guidance progression"):
                    m_year = re.match(r"(20\d{2})", first_txt)
                    if m_year and int(m_year.group(1)) < target_year:
                        insert_at = rr
                        break
                    if m_year and int(m_year.group(1)) > target_year:
                        _existing_start, _existing_header, existing_end = _block_bounds(first_txt)
                        fallback_insert = max(fallback_insert or 0, int(existing_end or rr) + 1)
                elif first_txt.endswith("open guidance") or first_txt.endswith("revisions") or first_txt == "Quarterly guidance timeline / revision log":
                    insert_at = fallback_insert if fallback_insert is not None else rr
                    break
            else:
                if fallback_insert is not None:
                    insert_at = fallback_insert
        ws.insert_rows(insert_at, 2 + len(rows))
        blue = PatternFill("solid", fgColor="5B9BD5")
        header_fill = PatternFill("solid", fgColor="EAF3FB")
        neutral = PatternFill("solid", fgColor="FFFFFF")
        neutral_alt = PatternFill("solid", fgColor="F6F9FC")
        border = Border(bottom=Side(style="thin", color="D9E2EF"))
        ws.merge_cells(start_row=insert_at, start_column=1, end_row=insert_at, end_column=max_col)
        for cc in range(1, max_col + 1):
            cell = ws.cell(insert_at, cc)
            cell.fill = blue
            cell.border = border
        title_cell = ws.cell(insert_at, 1, title)
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[insert_at].height = 22
        headers = ["Metric", "Initial guide", "Q1 update", "Q2 update", "Q3 update", "Q4 update", "Actual", "Status", "Notes/source"]
        header_row = insert_at + 1
        for cc in range(1, max_col + 1):
            value = headers[cc - 1] if cc <= len(headers) else ""
            cell = ws.cell(header_row, cc, value)
            cell.fill = header_fill
            cell.font = Font(bold=True, size=11)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        if max_col >= 9:
            try:
                ws.merge_cells(start_row=header_row, start_column=9, end_row=header_row, end_column=max_col)
            except Exception:
                pass
        ws.row_dimensions[header_row].height = 22

        def _status_fill_local(status: Any) -> PatternFill:
            low = str(status or "").strip().lower()
            if low in {"hit", "beat"}:
                return PatternFill("solid", fgColor="66C2A5")
            if low == "missed":
                return PatternFill("solid", fgColor="D55E00")
            if low == "completed":
                return PatternFill("solid", fgColor="009E73")
            if low == "open":
                return PatternFill("solid", fgColor="A6CEE3")
            return PatternFill("solid", fgColor="56B4E9") if low == "on track" else neutral

        row_idx = header_row + 1
        for item in rows:
            fill = neutral_alt if row_idx % 2 else neutral
            values = [
                item.get("metric"),
                item.get("initial"),
                item.get("q1"),
                item.get("q2"),
                item.get("q3"),
                item.get("q4"),
                item.get("actual"),
                item.get("status"),
                item.get("note"),
            ]
            for cc in range(1, max_col + 1):
                value = values[cc - 1] if cc <= len(values) else ""
                cell = ws.cell(row_idx, cc, value)
                cell.fill = _status_fill_local(value) if cc == 8 else fill
                cell.font = Font(size=11)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc >= 9)
            if max_col >= 9:
                try:
                    ws.merge_cells(start_row=row_idx, start_column=9, end_row=row_idx, end_column=max_col)
                except Exception:
                    pass
            ws.row_dimensions[row_idx].height = 24
            row_idx += 1

    def _append_quarter_note(quarter_label: str, category: str, note: str, metric: str) -> None:
        if "Quarter_Notes_UI" not in getattr(wb, "sheetnames", []):
            return
        qws = wb["Quarter_Notes_UI"]
        existing_blob = "\n".join(
            str(qws.cell(rr, cc).value or "")
            for rr in range(1, int(qws.max_row or 0) + 1)
            for cc in range(1, min(int(qws.max_column or 0), 4) + 1)
        )
        if note in existing_blob:
            return
        start = 0
        for rr in range(1, int(qws.max_row or 0) + 1):
            if str(qws.cell(rr, 1).value or "").strip() == quarter_label:
                start = rr
                break
        if not start:
            return
        end = int(qws.max_row or 0)
        for rr in range(start + 1, int(qws.max_row or 0) + 1):
            first = str(qws.cell(rr, 1).value or "").strip()
            if re.fullmatch(r"20\d{2}(?:-Q[1-4]|-\d{2}-\d{2})", first):
                end = rr - 1
                break
        insert_at = end + 1
        qws.insert_rows(insert_at, 1)
        template = max(start + 2, end)
        if template >= insert_at:
            template = start + 1
        for cc in range(1, min(int(qws.max_column or 0), 4) + 1):
            src = qws.cell(template, cc)
            dst = qws.cell(insert_at, cc)
            if src.has_style:
                dst._style = copy(src._style)
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
        qws.cell(insert_at, 1).value = None
        qws.cell(insert_at, 2).value = category
        qws.cell(insert_at, 3).value = note
        qws.cell(insert_at, 4).value = metric
        qws.row_dimensions[insert_at].height = qws.row_dimensions[template].height or 20.0

    def _rename_metric(old: str, new: str) -> None:
        for row in ws.iter_rows(min_row=1, max_row=int(ws.max_row or 0), min_col=1, max_col=max_col):
            for cell in row:
                if cell.value == old:
                    cell.value = new
                elif isinstance(cell.value, str) and old in cell.value and not cell.value.startswith("="):
                    if old == "EPS guidance" and new == "Adjusted EPS guidance":
                        txt = re.sub(r"(?<!Adjusted\s)\bEPS guidance\b", new, cell.value)
                    else:
                        txt = cell.value.replace(old, new)
                    while "Adjusted Adjusted" in txt:
                        txt = txt.replace("Adjusted Adjusted", "Adjusted")
                    cell.value = txt

    for row in ws.iter_rows(min_row=1, max_row=int(ws.max_row or 0), min_col=1, max_col=max_col):
        for cell in row:
            if isinstance(cell.value, str) and not cell.value.startswith("="):
                txt = cell.value
                while "Adjusted Adjusted" in txt:
                    txt = txt.replace("Adjusted Adjusted", "Adjusted")
                if txt != cell.value:
                    cell.value = txt

    if ticker_txt == "PBI":
        is_real_pbi_workbook = "PBI_Investment_Case" in getattr(wb, "sheetnames", [])
        if not is_real_pbi_workbook:
            _rename_metric("EPS guidance", "Adjusted EPS guidance")
            return
        pbi_section_labels = {
            str(ws.cell(rr, 1).value or "").strip().lower()
            for rr in range(1, int(ws.max_row or 0) + 1)
            if str(ws.cell(rr, 1).value or "").strip()
        }
        has_pbi_curated_layout = any(
            label.endswith("guidance progression")
            or label.endswith("open guidance")
            or label == "quarterly guidance timeline / revision log"
            for label in pbi_section_labels
        )
        _rename_metric("EPS guidance", "Adjusted EPS guidance")
        for year, rows in promise_source_support.pbi_guidance_progression_rows():
            _upsert_guidance_progression_section(year, rows)
        for rr, section, cols in _promise_rows():
            metric_col = cols.get("metric") or cols.get("milestone")
            metric_txt = str(ws.cell(rr, metric_col).value or "").strip() if metric_col else ""
            if metric_txt == "Adjusted EPS guidance":
                note_col = cols.get("source / note") or cols.get("notes/source")
                if note_col:
                    note_txt = str(ws.cell(rr, note_col).value or "")
                    note_txt = re.sub(r"(?<!Adjusted\s)\bEPS guidance\b", "Adjusted EPS guidance", note_txt)
                    while "Adjusted Adjusted" in note_txt:
                        note_txt = note_txt.replace("Adjusted Adjusted", "Adjusted")
                    ws.cell(rr, note_col).value = note_txt
            if section.endswith("guidance progression"):
                actual_col = cols.get("actual")
                status_col = cols.get("status")
                note_col = cols.get("notes/source")
                if metric_txt == "Adjusted EPS guidance" and actual_col:
                    ws.cell(rr, actual_col).value = "$1.35"
                    if status_col:
                        ws.cell(rr, status_col).value = "Hit"
                    if note_col:
                        ws.cell(rr, note_col).value = "2025 year adjusted diluted EPS."
                elif metric_txt == "FCF target" and actual_col:
                    ws.cell(rr, actual_col).value = "$358.3m"
                    if status_col:
                        ws.cell(rr, status_col).value = "Hit"
                    if note_col:
                        ws.cell(rr, note_col).value = "2025 year source-defined Free Cash Flow."
            if section.endswith("open guidance") and metric_txt == "Cost savings target":
                horizon_col = cols.get("horizon")
                note_col = cols.get("notes/source")
                if horizon_col:
                    ws.cell(rr, horizon_col).value = "Annualized program"
                if note_col:
                    ws.cell(rr, note_col).value = "Latest run-rate $157m; target $180m-$200m."

        if has_pbi_curated_layout:
            pbi_source_maps = promise_source_support.pbi_source_record_maps()
            pbi_source_revenue = pbi_source_maps["revenue"]
            pbi_source_fcf = pbi_source_maps["fcf"]
            pbi_source_adjusted_ebit = pbi_source_maps["adjusted_ebit"]
            pbi_source_adjusted_eps = pbi_source_maps["adjusted_eps"]
            _pbi_source_record_values = promise_source_support.pbi_source_record_values

            for rr, section, cols in _promise_rows():
                if not section.endswith("revisions"):
                    continue
                metric_col = cols.get("metric")
                change_col = cols.get("change type")
                stated_col = cols.get("stated in")
                horizon_col = cols.get("horizon")
                actual_col = cols.get("actual")
                progress_col = cols.get("progress / run-rate")
                status_col = cols.get("status")
                note_col = cols.get("source / note")
                if not (metric_col and stated_col and actual_col):
                    continue
                metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
                stated_txt = str(ws.cell(rr, stated_col).value or "").strip()
                horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip() if horizon_col else ""
                if metric_txt == "Revenue guidance" and stated_txt in pbi_source_revenue:
                    actual, progress, status, note, change_type = _pbi_source_record_values(pbi_source_revenue[stated_txt])
                    _set_promise_row_semantics(
                        ws,
                        rr,
                        cols,
                        change_type=change_type or None,
                        actual=actual,
                        progress=progress,
                        status=status,
                        note=note,
                    )
                elif metric_txt == "FCF target" and stated_txt in pbi_source_fcf:
                    actual, progress, status, note, change_type = _pbi_source_record_values(pbi_source_fcf[stated_txt])
                    annual_year = _promise_annual_year(horizon_txt)
                    stated_year, stated_q = _promise_stated_quarter_parts(stated_txt)
                    if progress_col and annual_year is not None and stated_year == annual_year and stated_q in {1, 2, 3}:
                        ws.cell(rr, actual_col).value = actual
                        ws.cell(rr, progress_col).value = progress
                    else:
                        ws.cell(rr, actual_col).value = actual
                        if progress_col:
                            ws.cell(rr, progress_col).value = progress
                    _set_promise_row_semantics(
                        ws,
                        rr,
                        cols,
                        change_type=change_type or None,
                        status=status,
                        note=note,
                    )
                elif metric_txt == "Adjusted EBIT guidance" and stated_txt in pbi_source_adjusted_ebit:
                    actual, progress, status, note, change_type = _pbi_source_record_values(pbi_source_adjusted_ebit[stated_txt])
                    _set_promise_row_semantics(
                        ws,
                        rr,
                        cols,
                        change_type=change_type or None,
                        actual=actual,
                        progress=progress,
                        status=status,
                        note=note,
                    )
                elif metric_txt == "Adjusted EPS guidance" and stated_txt in pbi_source_adjusted_eps:
                    actual, progress, status, note, change_type = _pbi_source_record_values(pbi_source_adjusted_eps[stated_txt])
                    _set_promise_row_semantics(
                        ws,
                        rr,
                        cols,
                        change_type=change_type or None,
                        actual=actual,
                        progress=progress,
                        status=status,
                        note=note,
                    )

            for pbi_2026_q1_row in promise_source_support.pbi_2026_q1_rows():
                _upsert_timeline_row("2026-Q1 revisions", pbi_2026_q1_row)

            for section_name, row_values in promise_source_support.pbi_cost_rows():
                _upsert_timeline_row(section_name, row_values)

            start, header, end = _block_bounds("2026 open guidance")
            if start and header and end >= header:
                open_cols = _header_map(header)
                for rr in range(header + 1, end + 1):
                    metric_txt = str(ws.cell(rr, open_cols.get("metric", 1)).value or "").strip()
                    if not metric_txt:
                        continue
                    guide_col = open_cols.get("current guide") or open_cols.get("new/current guide")
                    horizon_col = open_cols.get("horizon")
                    status_col = open_cols.get("status")
                    note_col = open_cols.get("notes/source") or open_cols.get("source / note")
                    if metric_txt == "Revenue guidance":
                        if guide_col:
                            ws.cell(rr, guide_col).value = "$1.8bn-$1.86bn"
                        if horizon_col:
                            ws.cell(rr, horizon_col).value = "2026 year"
                        if status_col:
                            ws.cell(rr, status_col).value = "Open"
                        if note_col:
                            ws.cell(rr, note_col).value = "2026 year Revenue guidance updated to $1.8bn-$1.86bn."
                    elif metric_txt in {"Adjusted EBIT guidance", "Adjusted EPS guidance", "FCF target"}:
                        if horizon_col:
                            ws.cell(rr, horizon_col).value = "2026 year"

            pbi_rows_to_delete: List[int] = []
            for rr, section, cols in _promise_rows():
                metric_col = cols.get("metric") or cols.get("milestone")
                if not metric_col:
                    continue
                metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
                if section.endswith("open guidance") and metric_txt == "Revenue guidance":
                    guide_col = cols.get("current guide") or cols.get("new/current guide")
                    horizon_col = cols.get("horizon")
                    status_col = cols.get("status")
                    note_col = cols.get("notes/source") or cols.get("source / note")
                    if guide_col:
                        ws.cell(rr, guide_col).value = "$1.8bn-$1.86bn"
                    if horizon_col:
                        ws.cell(rr, horizon_col).value = "2026 year"
                    if status_col:
                        ws.cell(rr, status_col).value = "Open"
                    if note_col:
                        ws.cell(rr, note_col).value = "2026 year Revenue guidance updated to $1.8bn-$1.86bn."
                if section == "2025-Q4 revisions" and metric_txt == "Revenue guidance":
                    new_col = cols.get("new/current guide")
                    horizon_col = cols.get("horizon")
                    new_txt = str(ws.cell(rr, new_col).value or "").strip() if new_col else ""
                    horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip() if horizon_col else ""
                    if "$425m-$465m" in new_txt or horizon_txt == "2025-Q4":
                        pbi_rows_to_delete.append(rr)
            for rr in sorted(set(pbi_rows_to_delete), reverse=True):
                ws.delete_rows(rr, 1)

            pbi_q4_promise_semantics = promise_source_support.pbi_q4_promise_semantics()
            for rr, section, cols in _promise_rows():
                metric_col = cols.get("metric") or cols.get("milestone")
                metric_txt = str(ws.cell(rr, metric_col).value or "").strip() if metric_col else ""
                row_semantics = pbi_q4_promise_semantics.get((section, metric_txt))
                if not row_semantics:
                    continue
                change_txt, actual_txt, progress_txt, status_txt, note_txt = row_semantics
                _set_promise_row_semantics(
                    ws,
                    rr,
                    cols,
                    change_type=change_txt,
                    actual=actual_txt,
                    progress=progress_txt,
                    status=status_txt,
                    note=note_txt,
                )

            for q, category, note, metric in promise_source_support.pbi_quarter_notes():
                _append_quarter_note(q, category, note, metric)

            _remove_pbi_duplicate_cost_savings_timeline_rows(ws)

    elif ticker_txt == "ANF":
        _rename_metric("Adjusted EPS / EPS", "Adjusted EPS")
        if "ANF_Investment_Case" not in getattr(wb, "sheetnames", []):
            return

        def _anf_actuals_by_year_from_workbook() -> Dict[int, Dict[str, float]]:
            out: Dict[int, Dict[str, float]] = {}
            if "History_Q" not in getattr(wb, "sheetnames", []):
                return out
            hist_ws = wb["History_Q"]
            headers = {
                re.sub(r"[^a-z0-9]+", "_", str(hist_ws.cell(1, cc).value or "").strip().lower()).strip("_"): cc
                for cc in range(1, int(hist_ws.max_column or 0) + 1)
            }
            q_col = headers.get("quarter")
            if not q_col:
                return out

            def _hist_num(value: Any) -> Optional[float]:
                num = pd.to_numeric(value, errors="coerce")
                if pd.isna(num):
                    return None
                val = float(num)
                return val if math.isfinite(val) else None

            for rr in range(2, int(hist_ws.max_row or 0) + 1):
                qd = _date_or_none(hist_ws.cell(rr, q_col).value)
                fiscal_year = _hist_num(hist_ws.cell(rr, headers.get("fiscal_year", 0)).value) if headers.get("fiscal_year") else None
                year = int(fiscal_year) if fiscal_year is not None else (qd.year if qd is not None else None)
                if year is None:
                    continue
                for key, source_col in (
                    ("revenue", headers.get("revenue")),
                    ("capex", headers.get("capex")),
                ):
                    if not source_col:
                        continue
                    val = _hist_num(hist_ws.cell(rr, source_col).value)
                    if val is None:
                        continue
                    out.setdefault(int(year), {})[key] = out.setdefault(int(year), {}).get(key, 0.0) + val
            return out

        actuals_by_year = _anf_actuals_by_year_from_workbook()

        def _anf_annual_value(year: int, key: str) -> Optional[float]:
            val = actuals_by_year.get(int(year), {}).get(key)
            num = pd.to_numeric(val, errors="coerce")
            return float(num) if pd.notna(num) else None

        def _anf_annual_sales_growth(year: int) -> str:
            cur = _anf_annual_value(year, "revenue")
            prev = _anf_annual_value(year - 1, "revenue")
            if cur is None or prev is None or abs(prev) < 1e-9:
                return ""
            return f"{((cur / prev) - 1.0) * 100:.1f}%"

        def _anf_annual_capex_actual(year: int) -> str:
            capex_val = _anf_annual_value(year, "capex")
            return _format_lookup_actual_value("Capex", "capex", capex_val) if capex_val is not None else ""

        for year, rows in promise_source_support.anf_guidance_progression_rows(
            annual_sales_growth=_anf_annual_sales_growth,
            annual_capex_actual=_anf_annual_capex_actual,
        ):
            _upsert_guidance_progression_section(year, rows)
        anf_quarter_eps_actuals = promise_source_support.anf_quarter_eps_actuals()
        anf_diluted_share_progress = promise_source_support.anf_diluted_share_progress()
        anf_final_q4_rows = promise_source_support.anf_final_q4_rows()
        q4_note = promise_source_support.anf_q4_note()
        q4_adjusted_eps_note = promise_source_support.anf_q4_adjusted_eps_note()
        q4_share_note = promise_source_support.anf_q4_share_note()
        for section_name, row_values in promise_source_support.anf_final_q4_timeline_rows():
            _upsert_timeline_row(section_name, row_values)
        q2_diluted_share_section, q2_diluted_share_row = promise_source_support.anf_q2_diluted_share_row()
        _upsert_timeline_row(
            q2_diluted_share_section,
            q2_diluted_share_row,
        )
        for rr, section, cols in _promise_rows():
            metric_col = cols.get("metric")
            actual_col = cols.get("actual")
            progress_col = cols.get("progress / run-rate")
            status_col = cols.get("status")
            note_col = cols.get("source / note") or cols.get("notes/source")
            horizon_col = cols.get("horizon")
            stated_col = cols.get("stated in")
            source_date_col = cols.get("source date")
            metric_txt = str(ws.cell(rr, metric_col).value or "").strip() if metric_col else ""
            horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip() if horizon_col else ""
            stated_txt = str(ws.cell(rr, stated_col).value or "").strip() if stated_col else ""
            if section == "2025-Q4 revisions" and horizon_txt == "2025 year" and metric_txt in anf_final_q4_rows:
                _prev_txt, _new_txt, actual_txt, progress_txt, status_txt = anf_final_q4_rows[metric_txt]
                note_txt = q4_note
                if metric_txt == "Adjusted EPS":
                    note_txt = q4_adjusted_eps_note
                elif metric_txt == "Diluted shares":
                    note_txt = q4_share_note
                _set_promise_row_semantics(
                    ws,
                    rr,
                    cols,
                    change_type="Completed",
                    actual=actual_txt,
                    progress=progress_txt,
                    status=status_txt,
                    note=note_txt,
                )
                if source_date_col:
                    ws.cell(rr, source_date_col).value = "2026-03-04"
                continue
            if metric_txt != "Adjusted EPS":
                if metric_txt == "Diluted shares" and stated_txt in anf_diluted_share_progress and horizon_txt == "2025 year":
                    actual_txt, progress_txt, status_txt, note_txt = anf_diluted_share_progress[stated_txt]
                    if actual_col:
                        ws.cell(rr, actual_col).value = actual_txt
                    if progress_col:
                        ws.cell(rr, progress_col).value = progress_txt
                    if status_col:
                        ws.cell(rr, status_col).value = status_txt
                    if note_col:
                        ws.cell(rr, note_col).value = note_txt
                continue
            if stated_txt in anf_quarter_eps_actuals and horizon_txt == "2025 year":
                actual_txt, progress_txt, status_txt, note_txt = anf_quarter_eps_actuals[stated_txt]
                if actual_col:
                    ws.cell(rr, actual_col).value = actual_txt
                if progress_col:
                    ws.cell(rr, progress_col).value = progress_txt
                if status_col:
                    ws.cell(rr, status_col).value = status_txt
                if note_col:
                    ws.cell(rr, note_col).value = note_txt
                continue
            if section.endswith("guidance progression") or (horizon_txt == "2025 year" and stated_txt == "2025-Q4"):
                if actual_col:
                    ws.cell(rr, actual_col).value = "$9.86 adjusted"
                if status_col:
                    ws.cell(rr, status_col).value = "Missed"
                if note_col:
                    ws.cell(rr, note_col).value = "GAAP EPS $10.46 also reported."
        _ensure_anf_promise_hidden_source_keys(ws)

    elif ticker_txt == "GPRE":
        for section_name, row_values in promise_source_support.gpre_source_rows():
            _upsert_timeline_row(section_name, row_values)
        for rr, section, cols in _promise_rows():
            metric_col = cols.get("metric") or cols.get("milestone")
            actual_col = cols.get("actual")
            progress_col = cols.get("progress / run-rate")
            status_col = cols.get("status")
            note_col = cols.get("source / note") or cols.get("notes/source")
            horizon_col = cols.get("horizon")
            stated_col = cols.get("stated in")
            source_date_col = cols.get("source date")
            metric_txt = str(ws.cell(rr, metric_col).value or "").strip() if metric_col else ""
            horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip() if horizon_col else ""
            stated_txt = str(ws.cell(rr, stated_col).value or "").strip() if stated_col else ""
            if metric_txt == "2026 year 45Z EBITDA guidance" and stated_txt == "2026-Q1" and horizon_txt == "2026 year":
                _set_promise_row_semantics(
                    ws,
                    rr,
                    cols,
                    actual="$55.2m",
                    progress="YTD: $55.2m",
                    status="On track",
                    note="2026-Q1 45Z contribution tracks annual guide; full-year horizon remains open.",
                )
            if metric_txt == "45Z monetization" and stated_txt == "2025-Q4" and horizon_txt == "2025-Q4":
                _set_promise_row_semantics(
                    ws,
                    rr,
                    cols,
                    actual="$23.4m",
                    progress="YTD: $49.9m",
                    status="Hit",
                    note="Q4 adjusted-EBITDA 45Z value was $23.4m; YTD adds Q3 $26.5m and Q4 $23.4m on the same value basis.",
                )
            if metric_txt == "Cost savings target" and horizon_txt == "Annualized program":
                if stated_txt == "2024-Q4":
                    if actual_col:
                        ws.cell(rr, actual_col).value = "$30m"
                    if progress_col:
                        ws.cell(rr, progress_col).value = "Executed: $30m"
                    if note_col:
                        ws.cell(rr, note_col).value = "Up to $50m identified; first $30m executed."
                elif stated_txt == "2025-Q1":
                    if actual_col:
                        ws.cell(rr, actual_col).value = "$45m"
                    if progress_col:
                        ws.cell(rr, progress_col).value = "Remaining: $5m"
                    if note_col:
                        ws.cell(rr, note_col).value = "Approximately $45m annualized savings accomplished; about $5m remaining."
                elif stated_txt == "2025-Q2":
                    if actual_col:
                        ws.cell(rr, actual_col).value = ">= $50m"
                    if progress_col:
                        ws.cell(rr, progress_col).value = "On pace to exceed $50m"
                    if status_col:
                        ws.cell(rr, status_col).value = "On track"
                    if note_col:
                        ws.cell(rr, note_col).value = (
                            "Cost reductions implemented in the first half; on pace to exceed the $50m annualized savings target."
                        )
            if metric_txt == "45Z facility qualification":
                if actual_col:
                    ws.cell(rr, actual_col).value = ""
                if progress_col:
                    ws.cell(rr, progress_col).value = "All 8 qualifying / AN operational"
                if status_col:
                    ws.cell(rr, status_col).value = "On track"
                if source_date_col:
                    ws.cell(rr, source_date_col).value = "2026-03-31"
                if note_col:
                    ws.cell(rr, note_col).value = "Conference metadata says all plants qualify from Jan. 1; Advantage Nebraska operational."

    for rr, _section, cols in _promise_rows():
        metric_col = cols.get("metric") or cols.get("milestone")
        horizon_col = cols.get("horizon")
        if not metric_col or not horizon_col:
            continue
        metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
        horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip()
        if not metric_txt or metric_txt.lower() in {"metric", "milestone"}:
            continue
        ws.cell(rr, 15).value = _promise_override_lifecycle_id(metric_txt, horizon_txt)
    ws.column_dimensions["O"].hidden = True
