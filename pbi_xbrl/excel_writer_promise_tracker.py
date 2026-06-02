"""Promise_Tracker_UI writer extracted from excel_writer_context."""
from __future__ import annotations

import datetime as dt
import hashlib
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Callable, Dict, List, Optional, Pattern, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .filing_evidence_shared import (
    build_canonical_subject_key as shared_build_canonical_subject_key,
    build_lifecycle_subject_key as shared_build_lifecycle_subject_key,
    build_parent_subject_key as shared_build_parent_subject_key,
    build_promise_lifecycle_key as shared_build_promise_lifecycle_key,
    classify_statement_evidence_role as shared_classify_statement_evidence_role,
    derive_lifecycle_state as shared_derive_lifecycle_state,
    derive_status_resolution_reason as shared_derive_status_resolution_reason,
    evidence_role as shared_evidence_role,
    infer_target_period_norm as shared_infer_target_period_norm,
    promise_candidate_drop_reason as shared_promise_candidate_drop_reason,
    qualify_promise_candidate as shared_qualify_promise_candidate,
    route_to_measurable_promise_candidate as shared_route_to_measurable_promise_candidate,
    source_class as shared_source_class,
    statement_class as shared_statement_class,
)
from .guidance_lexicon import (
    FORWARD_NOTES_LABEL,
    GUIDANCE_UI_METRIC_PRIORITY,
    extract_numeric_patterns as glx_extract_numeric_patterns,
    normalize_text as glx_normalize_text,
)
from .quarter_notes_lexicon import (
    compact_snippet as qn_compact_snippet,
    score_promise_candidate as qn_score_promise_candidate,
)


@dataclass
class PromiseTrackerWriterDeps:
    wb: Any
    promises: Optional[pd.DataFrame]
    slides_guidance: Optional[pd.DataFrame]
    promise_evidence_df: Optional[pd.DataFrame]
    ui_state: Dict[str, Any]
    ui_info_rows: List[Dict[str, Any]]
    company_profile: Any
    ticker: Any
    is_pbi_profile: bool
    is_gpre_profile: bool
    header_size: float
    apply_hyperlink_look: Callable[..., None]
    candidate_quality_key: Callable[..., Any]
    classify_pbi_metric_label: Callable[..., str]
    clean_target_bonus: Callable[..., Any]
    derive_split_target_meta: Callable[..., Dict[str, Any]]
    extract_45z_monetization_target_display: Callable[..., str]
    extract_money_targets_for_display: Callable[..., Any]
    extract_pbi_guidance_targets_multi: Callable[..., Any]
    extract_pbi_target_display: Callable[..., str]
    fmt_short_money_value_local: Callable[..., str]
    gpre_bad_visible_promise_reason: Callable[..., str]
    gpre_clean_visible_promise_metric: Callable[..., str]
    is_45z_crush_margin_support_only: Callable[..., bool]
    is_pbi_clean_sentence: Callable[..., bool]
    is_preferred_narrative_source: Callable[..., bool]
    load_profile_slide_signals: Callable[..., Any]
    looks_pbi_fragment_text: Callable[..., bool]
    management_theme_key: Callable[..., Any]
    pbi_promise_theme_re: Pattern[str]
    pbi_structured_guidance_items_for_qd: Callable[..., Any]
    pbi_structured_strategy_items_for_qd: Callable[..., Any]
    pbi_target_display_ok: Callable[..., bool]
    profile_slide_metric: Callable[..., str]
    promises_view: Callable[..., pd.DataFrame]
    resolve_col: Callable[..., Any]
    set_cell_comment: Callable[..., None]
    slide_signal_noise: Callable[..., bool]
    source_rank: Callable[..., int]
    split_target_group_key: Callable[..., Any]
    split_target_identity_key: Callable[..., Any]
    split_target_metric_display: Callable[..., str]
    split_target_scope_token: Callable[..., str]
    strong_45z_2026_target_display: Callable[..., str]
    text_fragment_penalty: Callable[..., Any]


def write_promise_tracker_ui_sheet(
    deps: PromiseTrackerWriterDeps,
    render_visible: bool = True,
) -> List[Dict[str, Any]]:
    wb = deps.wb
    promises = deps.promises
    slides_guidance = deps.slides_guidance
    promise_evidence_df = deps.promise_evidence_df
    ui_state = deps.ui_state
    ui_info_rows = deps.ui_info_rows
    company_profile = deps.company_profile
    ticker = deps.ticker
    is_pbi_profile = deps.is_pbi_profile
    is_gpre_profile = deps.is_gpre_profile
    header_size = deps.header_size
    _apply_hyperlink_look = deps.apply_hyperlink_look
    _candidate_quality_key = deps.candidate_quality_key
    _classify_pbi_metric_label = deps.classify_pbi_metric_label
    _clean_target_bonus = deps.clean_target_bonus
    _derive_split_target_meta = deps.derive_split_target_meta
    _extract_45z_monetization_target_display = deps.extract_45z_monetization_target_display
    _extract_money_targets_for_display = deps.extract_money_targets_for_display
    _extract_pbi_guidance_targets_multi = deps.extract_pbi_guidance_targets_multi
    _extract_pbi_target_display = deps.extract_pbi_target_display
    _fmt_short_money_value_local = deps.fmt_short_money_value_local
    _gpre_bad_visible_promise_reason = deps.gpre_bad_visible_promise_reason
    _gpre_clean_visible_promise_metric = deps.gpre_clean_visible_promise_metric
    _is_45z_crush_margin_support_only = deps.is_45z_crush_margin_support_only
    _is_pbi_clean_sentence = deps.is_pbi_clean_sentence
    _is_preferred_narrative_source = deps.is_preferred_narrative_source
    _load_profile_slide_signals = deps.load_profile_slide_signals
    _looks_pbi_fragment_text = deps.looks_pbi_fragment_text
    _management_theme_key = deps.management_theme_key
    _pbi_promise_theme_re = deps.pbi_promise_theme_re
    _pbi_structured_guidance_items_for_qd = deps.pbi_structured_guidance_items_for_qd
    _pbi_structured_strategy_items_for_qd = deps.pbi_structured_strategy_items_for_qd
    _pbi_target_display_ok = deps.pbi_target_display_ok
    _profile_slide_metric = deps.profile_slide_metric
    _promises_view = deps.promises_view
    _resolve_col = deps.resolve_col
    _set_cell_comment_local = deps.set_cell_comment
    _slide_signal_noise = deps.slide_signal_noise
    _source_rank = deps.source_rank
    _split_target_group_key = deps.split_target_group_key
    _split_target_identity_key = deps.split_target_identity_key
    _split_target_metric_display = deps.split_target_metric_display
    _split_target_scope_token = deps.split_target_scope_token
    _strong_45z_2026_target_display = deps.strong_45z_2026_target_display
    _text_fragment_penalty = deps.text_fragment_penalty

    ws = wb.create_sheet("Promise_Tracker_UI") if render_visible else None
    qa_rows: List[Dict[str, Any]] = []
    ts = datetime.now(dt.UTC).strftime("%Y-%m-%d %H:%M:%S UTC")
    quarter_note_rows_seed = ui_state.get("quarter_notes_ui_rows", {}) if isinstance(ui_state, dict) else {}
    has_qnote_recall_seed = any(bool(rows) for rows in quarter_note_rows_seed.values()) if isinstance(quarter_note_rows_seed, dict) else False

    def _store_tracker_state(grouped_rows: Dict[date, List[Dict[str, Any]]], q_window_local: List[date]) -> None:
        def _state_tracker_copy(item_in: Dict[str, Any]) -> Dict[str, Any]:
            out_item = dict(item_in)
            if is_pbi_profile:
                metric_label = _classify_pbi_metric_label(
                    " | ".join(
                        [
                            str(out_item.get("metric_display") or ""),
                            str(out_item.get("metric") or ""),
                            str(out_item.get("text_full") or out_item.get("text_snippet") or ""),
                            str(out_item.get("target_display") or out_item.get("target") or ""),
                        ]
                    ),
                    str(out_item.get("metric_display") or out_item.get("metric") or ""),
                )
                if metric_label:
                    out_item["metric_display"] = metric_label
                    if str(out_item.get("metric") or "").strip() in {"", "Revenue", "Adj EBIT", "Adj EPS", "FCF"}:
                        out_item["metric"] = metric_label
            return out_item

        ui_state["quarters"] = q_window_local
        ui_state["promise_tracker_rows_by_q"] = {
            qd_local: [_state_tracker_copy(it) for it in grouped_rows.get(qd_local, [])]
            for qd_local in q_window_local
        }
        promise_ids = [
            {"promise_id": str(it.get("promise_id") or "").strip()}
            for qd_local in q_window_local
            for it in grouped_rows.get(qd_local, [])
            if str(it.get("promise_id") or "").strip()
        ]
        if promise_ids:
            ui_state["promise_rows"] = pd.DataFrame(promise_ids).drop_duplicates(["promise_id"]).reset_index(drop=True)
        else:
            ui_state["promise_rows"] = pd.DataFrame(columns=["promise_id"])

    if (promises is None or promises.empty) and not has_qnote_recall_seed:
        _store_tracker_state({}, [])
        if render_visible and ws is not None:
            ws["A1"] = f"Generated at {ts} | Quarter list view"
            ws["A2"] = "No data."
            ws.freeze_panes = "A2"
        return qa_rows

    if promises is None or promises.empty:
        p = pd.DataFrame(
            [
                {
                    "promise_id": "",
                    "promise_text": "",
                    "metric_tag": "",
                    "created_quarter": None,
                    "last_seen_quarter": None,
                    "first_seen_evidence_quarter": None,
                    "last_seen_evidence_quarter": None,
                    "form": "",
                    "doc": "",
                    "section_or_page": "",
                    "source_type": "",
                    "source_evidence_json": "",
                    "target_kind": "",
                    "target_time": "",
                    "promise_type": "",
                }
            ]
        )
    else:
        p = _promises_view().copy()
    pid_col = _resolve_col(p, ["promise_id", "id"])
    txt_col = _resolve_col(p, ["promise_text", "statement", "claim", "evidence_snippet"])
    metric_col = _resolve_col(p, ["metric_tag", "metric"])
    created_col = _resolve_col(p, ["created_quarter", "created_q", "first_seen_q", "first_seen_quarter", "quarter"])
    last_seen_col = _resolve_col(p, ["last_seen_quarter", "last_seen_q"])
    first_seen_ev_col = _resolve_col(p, ["first_seen_evidence_quarter", "first_seen_quarter", "created_quarter"])
    last_seen_ev_col = _resolve_col(p, ["last_seen_evidence_quarter", "last_seen_quarter", "created_quarter"])
    form_col = _resolve_col(p, ["form"])
    doc_col = _resolve_col(p, ["doc", "doc_path", "source_doc"])
    section_col = _resolve_col(p, ["section_or_page", "section", "page"])
    method_col = _resolve_col(p, ["method", "source_type"])
    ev_json_col = _resolve_col(p, ["source_evidence_json", "evidence_history_json", "evidence_json"])
    target_kind_col = _resolve_col(p, ["target_kind"])
    deadline_col = _resolve_col(p, ["target_time", "deadline"])
    promise_type_col = _resolve_col(p, ["promise_type"])
    if pid_col is None or txt_col is None:
        _store_tracker_state({}, [])
        if render_visible and ws is not None:
            ws["A1"] = f"Generated at {ts} | Quarter list view"
            ws["A2"] = "Missing required source columns."
            ws.freeze_panes = "A2"
        return qa_rows

    def _parse_json(raw: Any) -> Dict[str, Any]:
        if isinstance(raw, dict):
            return raw
        if isinstance(raw, list) and raw and isinstance(raw[0], dict):
            return raw[0]
        if not isinstance(raw, str) or not raw.strip():
            return {}
        try:
            z = json.loads(raw)
            if isinstance(z, dict):
                return z
            if isinstance(z, list) and z and isinstance(z[0], dict):
                return z[0]
        except Exception:
            return {}
        return {}

    def _qend(x: Any) -> Optional[date]:
        t = pd.to_datetime(x, errors="coerce")
        if pd.isna(t):
            return None
        return pd.Timestamp(t).to_period("Q").end_time.date()

    def _to_ts(x: Any) -> pd.Timestamp:
        t = pd.to_datetime(x, errors="coerce")
        return pd.Timestamp(t) if pd.notna(t) else pd.Timestamp("1900-01-01")

    def _quarter_lbl(v: Any) -> str:
        t = pd.to_datetime(v, errors="coerce")
        if pd.isna(t):
            return "N/A"
        qn = ((int(t.month) - 1) // 3) + 1
        return f"Q{qn} {int(t.year)}"

    def _source_meta(row: pd.Series) -> Dict[str, Any]:
        ev = _parse_json(row.get(ev_json_col) if ev_json_col else None)
        return {
            "source_type": str(row.get(method_col) or ev.get("doc_type") or ev.get("source_type") or "promise_tracker_ui"),
            "form": str(row.get(form_col) or ev.get("form") or ""),
            "accn": str(row.get("accn") or ev.get("accn") or ""),
            "filed": row.get("filed") or ev.get("filed") or None,
            "doc": str(row.get(doc_col) or ev.get("doc_path") or ev.get("doc_name") or ""),
            "section": str(row.get(section_col) or ev.get("section_or_page") or ev.get("section") or ""),
            "snippet": str(row.get("evidence_snippet") or ev.get("snippet") or ""),
        }

    def _clean_metric(raw_metric: Any) -> str:
        m = str(raw_metric or "").strip()
        if m.lower() in {"", "nan", "none", "n/a"}:
            return ""
        return m

    q_ui = ui_state.get("quarters") or []
    q_window: List[date] = []
    promise_progress_quarter_window = 16
    tmp_q: List[date] = []
    for _, r in p.iterrows():
        qd = _qend(r.get(created_col) if created_col else None) or _qend(r.get(last_seen_col) if last_seen_col else None)
        if qd is not None:
            tmp_q.append(qd)
    if q_ui:
        q_window = sorted(
            {q for q in [*q_ui, *tmp_q] if isinstance(q, date)},
            reverse=True,
        )[:promise_progress_quarter_window]
    else:
        q_window = sorted(set(tmp_q), reverse=True)[:promise_progress_quarter_window]
    q_set = set(q_window)

    metric_priority = {m: i for i, m in enumerate(GUIDANCE_UI_METRIC_PRIORITY)}
    metric_priority["Tone / expectations"] = 98
    metric_priority[FORWARD_NOTES_LABEL] = 99
    metric_priority["Strategic milestone"] = 6
    promise_metric_allow = {
        "Revenue",
        "Revenue guidance",
        "Adj EBITDA",
        "Adj EBIT",
        "Adjusted EBIT",
        "Adjusted EBIT guidance",
        "Adj EPS",
        "EPS guidance",
        "FCF",
        "FCF target",
        "Capex",
        "Cost savings",
        "Cost savings target",
        "Restructuring charges",
        "Net debt / leverage",
        "Deleveraging target",
        "Capital allocation",
        "PB Bank liquidity release",
        "Strategic milestone",
        "45Z monetization / EBITDA",
        "Debt reduction",
        "SendTech / Presort operating target",
        "Management target",
    }
    promise_priority_terms = tuple(
        str(t).strip().lower()
        for t in list(getattr(company_profile, "promise_priority_terms", ()) or [])
        if str(t).strip()
    )
    guidance_anchor_re = re.compile(
        r"\b(guidance|outlook|target|targets|range|between|expect|expects|plan|intend|on track|forecast|full[ -]?year|next quarter|next fiscal year|reaffirm)\b",
        re.I,
    )
    milestone_completion_re = re.compile(
        r"\b(completed|fully operational|online|executed|sale completed|started up|commercial operation|closed)\b",
        re.I,
    )
    milestone_progress_re = re.compile(
        r"\b(on track|ramping|under construction|progressing|began|begin|advance(?:d|s)?|commissioning)\b",
        re.I,
    )
    preferred_promise_source_re = re.compile(
        r"(earnings_release|press_release|presentation|slides|transcript|ceo|shareholder|mda|management discussion)",
        re.I,
    )

    def _promise_theme_key(metric_name: str, txt_in: str, period_key_in: str) -> str:
        return _management_theme_key(metric_name, txt_in, period_key_in)

    def _promise_target_strength_key(item: Dict[str, Any]) -> int:
        metric_name = str(item.get("metric") or item.get("metric_ref") or "").strip()
        txt_full = str(item.get("text_full") or item.get("text_snippet") or "")
        q_hint = item.get("quarter") or item.get("first_seen_quarter_end") or item.get("last_seen_quarter_end")
        target_txt = item.get("target_display") or item.get("target") or ""
        if str(metric_name).strip().lower() != "45z monetization / ebitda":
            return 0
        strong_target = _strong_45z_2026_target_display(txt_full, q_hint, target_txt)
        if not strong_target:
            return 0
        target_amounts = _extract_money_targets_for_display(strong_target)
        if not target_amounts:
            return 0
        return int(round(float(max(target_amounts)) / 1e6))

    def _promise_quality_key(item: Dict[str, Any]) -> Tuple[int, int, int, int, float, int]:
        src = dict(item.get("source") or {})
        base_key = _candidate_quality_key(
            item.get("text_full"),
            src.get("source_type"),
            src.get("doc"),
            item.get("doc_priority"),
            item.get("score"),
        )
        return base_key[:1] + (
            -_promise_target_strength_key(item),
        ) + base_key[1:]

    def _map_raw_metric(raw_metric: str) -> str:
        m = str(raw_metric or "").lower()
        if m == "":
            return ""
        if "45z" in m or "tax credit" in m or "monetization" in m:
            return "45Z monetization / EBITDA"
        if "revenue" in m or "sales" in m or "top line" in m:
            return "Revenue"
        if "ebitda" in m:
            return "Adj EBITDA"
        if "eps" in m or "earnings per share" in m:
            return "Adj EPS"
        if "fcf" in m or "free cash flow" in m:
            return "FCF"
        if "capex" in m or "capital expenditure" in m:
            return "Capex"
        if "saving" in m:
            return "Cost savings"
        if "restructur" in m or "one-time" in m:
            return "Restructuring charges"
        if "lever" in m or "net debt" in m or "debt" in m:
            return "Net debt / leverage"
        if "buyback" in m or "repurchase" in m or "dividend" in m or "capital allocation" in m:
            return "Capital allocation"
        if "milestone" in m or "initiative" in m:
            return "Milestone"
        return ""

    def _guidance_period_label_from_norm(pnorm: str, asof_q: date) -> str:
        p = str(pnorm or "").strip()
        if not p or p == "UNK":
            return ""
        if p == "FY+1":
            return f"FY {int(asof_q.year) + 1}"
        m_fy = re.match(r"FY(20\d{2})$", p)
        if m_fy:
            return f"FY {int(m_fy.group(1))}"
        m_q = re.match(r"Q(20\d{2})Q([1-4])$", p)
        if m_q:
            return f"Q{int(m_q.group(2))} {int(m_q.group(1))}"
        return p

    def _guidance_value_snip(it: Dict[str, Any]) -> str:
        kind = str(it.get("kind") or "")
        unit = str(it.get("unit") or "")
        if kind == "range" and it.get("low") is not None and it.get("high") is not None:
            lo = float(it.get("low"))
            hi = float(it.get("high"))
            if unit == "$m":
                return f"${lo/1e6:,.1f}m-${hi/1e6:,.1f}m"
            if unit == "$":
                return f"${lo:,.2f}-${hi:,.2f}"
            if unit == "%":
                return f"{lo:.1f}% - {hi:.1f}%"
            if unit == "x":
                return f"{lo:.2f}x-{hi:.2f}x"
            return f"{lo:,.2f}-{hi:,.2f}"
        if kind == "point" and it.get("value") is not None:
            v = float(it.get("value"))
            if unit == "$m":
                return f"${v/1e6:,.1f}m"
            if unit == "$":
                return f"${v:,.2f}"
            if unit == "%":
                return f"{v:.1f}%"
            if unit == "x":
                return f"{v:.2f}x"
            return f"{v:,.2f}"
        return qn_compact_snippet(str(it.get("text") or ""), 240)

    def _has_structured_numeric_target(text_in: str, require_money_like: bool = False) -> bool:
        txt_local = str(text_in or "")
        for nh in glx_extract_numeric_patterns(txt_local):
            if not isinstance(nh, dict):
                continue
            k = str(nh.get("kind") or "").lower()
            unit = str(nh.get("unit") or "").strip().lower()
            if k == "range" and (nh.get("value_low") is not None or nh.get("value_high") is not None):
                if not require_money_like:
                    return True
                if unit in {"$m", "$", "%", "bps", "x"}:
                    return True
            if k == "point" and nh.get("value_point") is not None:
                if not require_money_like:
                    return True
                if unit in {"$m", "$", "%", "bps", "x"}:
                    return True
        if require_money_like and re.search(r"\$\s*[0-9]", txt_local) and re.search(r"\b(?:to|through|between|[-–—])\b", txt_local, re.I):
            return True
        return False

    def _is_guidance_header_line(text_in: str) -> bool:
        t = str(text_in or "").lower()
        if not re.search(r"\b(low|high)\b", t):
            return False
        metric_hits = 0
        for kw in ("revenue", "adjusted ebit", "adjusted ebitda", "adjusted eps", "free cash flow", "fcf", "capex"):
            if kw in t:
                metric_hits += 1
        if metric_hits >= 2:
            return True
        if re.search(r"\b(provides?\s+the\s+following\s+guidance|guidance ranges?\s+are)\b", t):
            return True
        return False

    winners: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    milestone_suppressed_count = 0
    historical_results_re = re.compile(
        r"\b(compared\s+(?:with|to)|vs\.?|versus)\b.*\b(primarily due to|as a result of|partially offset|"
        r"increased|decreased|declined|grew|improved|compressed|expanded)\b",
        re.I,
    )
    market_backdrop_re = re.compile(
        r"\b(new vehicles? sales?|vehicles? during\s+20\d{2}|represented approximately .*% of new vehicles? sales|market share)\b",
        re.I,
    )
    fallback_allow_numeric_re = re.compile(
        r"\b(annualized cost reductions?|additional annualized savings|one-time charge|approximately\s+\$[0-9])\b",
        re.I,
    )

    def _promise_guidance_type(metric_name: str, txt_in: str, period_key_in: str) -> str:
        low = str(txt_in or "").lower()
        if metric_name == "Net debt / leverage":
            return "ratio"
        if re.search(r"\b(annualized|annualised|run[- ]?rate)\b", low):
            return "run-rate"
        if re.search(r"\bone[- ]time\b[^.]{0,24}\b(charge|charges|cost|costs)\b", low):
            return "one-time"
        if re.search(r"\b(remainder|over\s+the?\s*next\s+year|into\s+20\d{2}|through\s+20\d{2})\b", low):
            if not re.search(r"\b(?:fy|fiscal)\s*(?:20\d{2}|\d{2})\b", low):
                return "ongoing"
        if str(period_key_in or "") not in {"", "UNK", "TIME_ANCHOR"}:
            return "period"
        return "text"

    for _, r in p.iterrows():
        pid = str(r.get(pid_col) or "").strip()
        if not pid:
            continue
        raw_txt = str(r.get(txt_col) or "")
        txt = glx_normalize_text(re.sub(r"\s*created\s+\d{4}-\d{2}-\d{2}\s*$", "", raw_txt, flags=re.I))
        if txt == "":
            ui_info_rows.append({"quarter": _qend(r.get(created_col) if created_col else None), "metric": "Promise_Tracker_UI", "severity": "info", "message": "dropped_reason=empty_text", "source": ""})
            continue
        if _is_guidance_header_line(txt):
            ui_info_rows.append({"quarter": _qend(r.get(created_col) if created_col else None), "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=guidance_header_noise | pid={pid}", "source": str(r.get(doc_col) if doc_col else '')})
            continue

        created_q = _qend(r.get(created_col) if created_col else None)
        first_seen_ev_q = _qend(r.get(first_seen_ev_col) if first_seen_ev_col else None) or created_q
        last_seen_ev_q = _qend(r.get(last_seen_ev_col) if last_seen_ev_col else None) or _qend(r.get(last_seen_col) if last_seen_col else None) or first_seen_ev_q
        display_q = last_seen_ev_q or first_seen_ev_q or created_q
        if display_q is None:
            ui_info_rows.append({"quarter": None, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=no_quarter | pid={pid}", "source": ""})
            continue
        if q_set and display_q not in q_set:
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=outside_quarter_window | pid={pid}", "source": ""})
            continue

        src = _source_meta(r)
        score_info = qn_score_promise_candidate(
            text=txt,
            source_type=str(src.get("source_type") or ""),
            form=str(src.get("form") or ""),
            doc_name=str(src.get("doc") or ""),
            section=str(src.get("section") or ""),
        )
        score_val = float(score_info.get("score") or 0.0)
        if bool(score_info.get("hard_exclude")):
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=boilerplate | pid={pid}", "source": str(src.get("doc") or "")})
            continue
        src_blob = " ".join(
            [
                str(src.get("source_type") or ""),
                str(src.get("doc") or ""),
                str(src.get("form") or ""),
                str(src.get("section") or ""),
            ]
        )
        preferred_promise_source = bool(preferred_promise_source_re.search(src_blob))
        priority_term_hits = sum(1 for term in promise_priority_terms if term and term in txt.lower())
        if preferred_promise_source:
            score_val += 6.0
        if priority_term_hits > 0:
            score_val += min(18.0, float(priority_term_hits) * 4.0)
        fragment_penalty = _text_fragment_penalty(txt)
        clean_target_bonus = _clean_target_bonus(txt)
        score_val -= float(fragment_penalty * 4.0)
        score_val += float(clean_target_bonus)
        promise_score_floor = 28.0 if preferred_promise_source else 30.0
        if score_val < promise_score_floor:
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=low_score | score={score_val:.1f} | pid={pid}", "source": str(src.get("doc") or "")})
            continue

        metric_txt = _clean_metric(r.get(metric_col) if metric_col else "")
        metric_from_text = str(score_info.get("metric_canon") or "")
        metric_from_profile = _profile_slide_metric(txt)
        metric_from_raw = _map_raw_metric(metric_txt)
        metric_canon = (
            metric_from_text
            if metric_from_text in promise_metric_allow
            else metric_from_profile
            if metric_from_profile in promise_metric_allow
            else metric_from_raw
        )
        if re.search(r"\b(net leverage|leverage ratio|deleverag|debt\/ebitda|net debt)\b", txt, re.I) and re.search(r"\b\d+(?:\.\d+)?\s*x\b", txt, re.I):
            metric_canon = "Net debt / leverage"
        if "45z" in txt.lower() and any(k in txt.lower() for k in ("ebitda", "monetization", "tax credit", "credits", "generation", "opportunity")):
            metric_canon = "45Z monetization / EBITDA"

        target_q = _qend(r.get(deadline_col) if deadline_col else None)
        target_kind = str(r.get(target_kind_col) or "").strip().lower() if target_kind_col else ""
        promise_type = str(r.get(promise_type_col) or "operational").strip().lower() if promise_type_col else "operational"
        has_numeric_range_or_point = _has_structured_numeric_target(txt, require_money_like=False)
        has_numeric = has_numeric_range_or_point
        has_time_anchor = bool(score_info.get("has_time_anchor")) or (target_q is not None)
        measurable = bool(score_info.get("measurable"))
        has_intent = bool(score_info.get("has_intent"))
        has_guidance_anchor = bool(guidance_anchor_re.search(txt))
        historical_result_only = bool(
            historical_results_re.search(txt)
            and not re.search(
                r"\b(target|targets|targeting|expect|expects|guidance|outlook|plan|plans|annualized|agreement|sale agreement|one-time charge)\b",
                txt,
                re.I,
            )
        )
        legal_hr_noise = bool(
            re.search(
                r"\b(one[- ]time payment|eligible to|employment agreement|base salary|target bonus|restricted stock|equity award|relocation)\b",
                txt,
                re.I,
            )
        )
        milestone_intent = bool(
            re.search(
                r"\b(complete|completed|launch|launched|close|closed|finish|finalize|execute|implement|deliver)\b",
                txt,
                re.I,
            )
        )
        if promise_type not in {"operational", "guidance_range", "milestone"}:
            promise_type = "operational"
        if promise_type == "operational" and (not has_numeric_range_or_point) and milestone_intent and has_time_anchor:
            promise_type = "milestone"
            if metric_canon in {"", "Other", "Unknown"}:
                metric_canon = "Strategic milestone"
            if not target_kind:
                target_kind = "milestone_due"
        is_milestone = promise_type == "milestone" or metric_canon in {"Milestone", "Strategic milestone"}
        if is_milestone:
            promise_type = "milestone"
            metric_canon = "Strategic milestone"
        if legal_hr_noise:
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=legal_or_hr_boilerplate | pid={pid}", "source": str(src.get('doc') or '')})
            continue
        if market_backdrop_re.search(txt):
            ui_info_rows.append(
                {
                    "quarter": display_q,
                    "metric": "Promise_Tracker_UI",
                    "severity": "info",
                    "message": f"dropped_reason=market_backdrop_not_promise | pid={pid}",
                    "source": str(src.get("doc") or ""),
                }
            )
            continue
        if historical_result_only:
            ui_info_rows.append(
                {
                    "quarter": display_q,
                    "metric": "Promise_Tracker_UI",
                    "severity": "info",
                    "message": f"dropped_reason=historical_result_commentary | pid={pid}",
                    "source": str(src.get("doc") or ""),
                }
            )
            continue

        if not has_intent:
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=no_intent | pid={pid}", "source": str(src.get("doc") or "")})
            continue
        # Revenue promises must be numeric + period anchored. Text-only forward language
        # belongs in tone/expectations to avoid misleading [Revenue] rows.
        revenue_numeric_ok = _has_structured_numeric_target(txt, require_money_like=True)
        if metric_canon == "Revenue" and not (revenue_numeric_ok and has_time_anchor):
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=revenue_without_numeric_anchor | pid={pid}", "source": str(src.get('doc') or '')})
            continue
        if not (has_guidance_anchor or has_time_anchor):
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=no_time_anchor | pid={pid}", "source": str(src.get("doc") or "")})
            continue
        if not preferred_promise_source and priority_term_hits <= 0 and promise_type != "guidance_range":
            ui_info_rows.append(
                {
                    "quarter": display_q,
                    "metric": "Promise_Tracker_UI",
                    "severity": "info",
                    "message": f"dropped_reason=low_priority_source_for_promise | pid={pid}",
                    "source": str(src.get("doc") or ""),
                }
            )
            continue
        if promise_type == "operational" and not target_kind and metric_canon not in {"", "Other", "Unknown"} and has_numeric_range_or_point and has_time_anchor:
            target_kind = "target"
        if promise_type == "operational" and not target_kind:
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=no_target_kind | pid={pid}", "source": str(src.get('doc') or '')})
            continue
        if promise_type == "operational" and target_q is None and str(score_info.get("period_key") or "") in {"", "UNK"}:
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=no_deadline_or_period | pid={pid}", "source": str(src.get('doc') or '')})
            continue
        if promise_type == "operational" and not has_numeric_range_or_point:
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=no_numeric_target_operational | pid={pid}", "source": str(src.get('doc') or '')})
            continue
        if not (has_numeric_range_or_point or promise_type == "milestone"):
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=no_metric_target | pid={pid}", "source": str(src.get("doc") or "")})
            continue
        if metric_canon in {"", "Other", "Unknown"} and not (measurable and has_time_anchor):
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=no_metric | pid={pid}", "source": str(src.get("doc") or "")})
            continue
        if metric_canon in {"", "Other", "Unknown"}:
            metric_canon = "Strategic milestone" if promise_type == "milestone" else ""
        if metric_canon in {"", FORWARD_NOTES_LABEL, "Tone / expectations"}:
            ui_info_rows.append({"quarter": display_q, "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=routed_to_narrative_layer | pid={pid}", "source": str(src.get('doc') or '')})
            continue

        period_label = str(score_info.get("period_label") or "")
        period_key = str(score_info.get("period_key") or "")
        if target_q is not None:
            tq = pd.Timestamp(target_q)
            q_num = int(tq.quarter)
            period_label = f"Q{q_num} {tq.year}"
            period_key = f"Q{tq.year}Q{q_num}"
        if period_key in {"", "UNK"} and has_time_anchor:
            period_key = "TIME_ANCHOR"
            period_label = period_label or "Time anchor"
        if metric_canon == "Revenue" and re.search(r"\brevenue\s+in\s+the\s+quarter\s+was\b", txt, re.I):
            ui_info_rows.append(
                {
                    "quarter": display_q,
                    "metric": "Promise_Tracker_UI",
                    "severity": "info",
                    "message": f"dropped_reason=quarter_actual_not_revenue_promise | pid={pid}",
                    "source": str(src.get("doc") or ""),
                }
            )
            continue
        if metric_canon == "Revenue":
            stale_year: Optional[int] = None
            m_fy = re.match(r"FY(20\\d{2})$", str(period_key))
            m_qy = re.match(r"Q(20\\d{2})Q([1-4])$", str(period_key))
            if m_fy:
                stale_year = int(m_fy.group(1))
            elif m_qy:
                stale_year = int(m_qy.group(1))
            elif target_q is not None:
                stale_year = int(target_q.year)
            if stale_year is not None and stale_year < int(display_q.year):
                ui_info_rows.append(
                    {
                        "quarter": display_q,
                        "metric": "Promise_Tracker_UI",
                        "severity": "info",
                        "message": f"dropped_reason=stale_revenue_period({stale_year}) | pid={pid}",
                        "source": str(src.get("doc") or ""),
                    }
                )
                continue

        source_date = _to_ts(src.get("filed") or r.get(last_seen_col) or r.get(created_col))
        guidance_type = _promise_guidance_type(metric_canon, txt, period_key)
        promise_theme_key = _promise_theme_key(metric_canon, txt, period_key)
        split_meta = _derive_split_target_meta(
            metric_canon,
            txt,
            period_key,
            display_q,
            src.get("source_type"),
            src.get("doc"),
            src.get("section"),
        )
        cand = {
            "promise_id": pid,
            "metric": metric_canon,
            "period_label": period_label,
            "period_key": period_key,
            "quarter": display_q,
            "text_full": txt,
            "text_snippet": qn_compact_snippet(txt, 240),
            "score": score_val,
            "source": src,
            "source_date": source_date,
            "doc_priority": int(score_info.get("doc_priority") or 0),
            "reasons": list(score_info.get("reasons") or []),
            "has_numeric": has_numeric,
            "has_time_anchor": has_time_anchor,
            "guidance_type": guidance_type,
            "as_of_quarter_end": str(display_q),
            "source_doc_end": str(display_q),
            "source_filed_date": pd.to_datetime(src.get("filed"), errors="coerce"),
            "first_seen_quarter_end": str(first_seen_ev_q) if first_seen_ev_q is not None else str(display_q),
            "last_seen_quarter_end": str(last_seen_ev_q) if last_seen_ev_q is not None else str(display_q),
            "referenced_years": sorted({int(y) for y in re.findall(r"(?<!\d)(20\d{2})(?!\d)", txt)}),
            "has_forward_intent": has_intent,
            "has_period_anchor": has_time_anchor,
            "target_period_norm": str(period_key or ""),
            "promise_type": promise_type,
            "theme_key": promise_theme_key,
            "target_display": _extract_45z_monetization_target_display(txt, display_q) if metric_canon == "45Z monetization / EBITDA" else "",
            "_fragment_penalty": fragment_penalty,
            "_clean_target_bonus": clean_target_bonus,
        }
        cand.update(split_meta)
        cand["promise_group"] = str(cand.get("target_group_key") or "")
        cand["metric_display"] = _split_target_metric_display(metric_canon, txt, cand)
        dedup_key = _split_target_identity_key(cand, metric_canon, period_key, display_q)

        prev = winners.get(dedup_key)
        if prev is None:
            winners[dedup_key] = cand
        else:
            better = False
            cand_qk = _promise_quality_key(cand)
            prev_qk = _promise_quality_key(prev)
            if cand_qk != prev_qk:
                better = cand_qk < prev_qk
            elif source_date != prev["source_date"]:
                better = source_date > prev["source_date"]
            else:
                better = len(str(cand["text_full"])) > len(str(prev["text_full"]))
            if better:
                ui_info_rows.append({"quarter": prev.get("quarter"), "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=duplicate_older | pid={prev.get('promise_id')}", "source": str(dict(prev.get('source') or {}).get('doc') or "")})
                winners[dedup_key] = cand
            else:
                ui_info_rows.append({"quarter": cand.get("quarter"), "metric": "Promise_Tracker_UI", "severity": "info", "message": f"dropped_reason=duplicate_older | pid={cand.get('promise_id')}", "source": str(src.get("doc") or "")})

    # Guidance-as-promises: split guidance snapshot rows into canonical per-metric items.
    gstore = ui_state.get("guidance_snapshot_by_q", {}) if isinstance(ui_state, dict) else {}
    if isinstance(gstore, dict) and gstore:
        for q_raw, items_raw in gstore.items():
            qd = _qend(q_raw)
            if qd is None:
                continue
            if q_set and qd not in q_set:
                continue
            if not isinstance(items_raw, list):
                continue
            for it in items_raw:
                if not isinstance(it, dict):
                    continue
                metric_name = str(it.get("metric") or "").strip()
                if metric_name in {"", FORWARD_NOTES_LABEL, "Other", "Unknown"}:
                    continue
                if metric_name not in promise_metric_allow:
                    continue
                kind = str(it.get("kind") or "").strip().lower()
                has_numeric_item = bool(
                    (kind == "range" and it.get("low") is not None and it.get("high") is not None)
                    or (kind == "point" and it.get("value") is not None)
                )
                period_norm = str(it.get("target_period_norm") or it.get("period_norm") or "UNK").strip() or "UNK"
                has_period_anchor_item = period_norm != "UNK"
                if not has_numeric_item or not has_period_anchor_item:
                    continue
                guidance_type = str(it.get("guidance_type") or "period")
                period_label = _guidance_period_label_from_norm(period_norm, qd)
                val_txt = _guidance_value_snip(it)
                raw_text_full = glx_normalize_text(str(it.get("text") or "").strip())
                if is_pbi_profile:
                    metric_label = _classify_pbi_metric_label(metric_name, metric_name) or metric_name
                    text_full = f"{period_label} {metric_label} {val_txt}".strip()
                else:
                    text_full = raw_text_full
                    if text_full:
                        text_full = f"{period_label} {val_txt} | {text_full}".strip()
                    else:
                        text_full = f"{period_label} {val_txt}".strip()
                src_it = dict(it.get("source") or {})
                filed_hint = src_it.get("filed") or src_it.get("source_filed_date") or q_raw
                source_date = _to_ts(filed_hint)
                promise_id = f"guidance:{qd.isoformat()}:{metric_name}:{period_norm}:{guidance_type}:{kind}"
                cand = {
                    "promise_id": promise_id,
                    "metric": metric_name,
                    "period_label": period_label,
                    "period_key": period_norm,
                    "quarter": qd,
                    "text_full": text_full,
                    "text_snippet": qn_compact_snippet(text_full, 260),
                    "score": float(it.get("score") or 80.0),
                    "source": {
                        "source_type": src_it.get("source_type") or "guidance_snapshot",
                        "form": src_it.get("form") or "",
                        "accn": src_it.get("accn") or "",
                        "filed": src_it.get("filed") or "",
                        "doc": src_it.get("doc") or "",
                        "section": src_it.get("section") or src_it.get("section_or_page") or "",
                        "snippet": raw_text_full or text_full,
                    },
                    "source_date": source_date,
                    "doc_priority": int(it.get("source_priority") or 0),
                    "reasons": ["guidance_snapshot_split"],
                    "has_numeric": True,
                    "has_time_anchor": True,
                    "guidance_type": guidance_type,
                    "as_of_quarter_end": str(qd),
                    "source_doc_end": str(q_raw),
                    "source_filed_date": pd.to_datetime(filed_hint, errors="coerce"),
                    "first_seen_quarter_end": str(it.get("first_seen_quarter_end") or qd),
                    "last_seen_quarter_end": str(it.get("last_seen_quarter_end") or qd),
                    "referenced_years": sorted({int(y) for y in re.findall(r"(?<!\\d)(20\\d{2})(?!\\d)", text_full)}),
                    "has_forward_intent": True,
                    "has_period_anchor": True,
                    "target_period_norm": period_norm,
                    "promise_type": "guidance_range",
                    "target_display": val_txt,
                    "text_full_raw": raw_text_full,
                }
                cand.update(
                    _derive_split_target_meta(
                        metric_name,
                        text_full,
                        period_norm,
                        qd,
                        src_it.get("source_type") or "guidance_snapshot",
                        src_it.get("doc") or "",
                        src_it.get("section") or src_it.get("section_or_page") or "",
                    )
                )
                cand["promise_group"] = str(cand.get("target_group_key") or "")
                cand["metric_display"] = _split_target_metric_display(metric_name, text_full, cand)
                dedup_key = _split_target_identity_key(cand, metric_name, period_norm, qd)
                prev = winners.get(dedup_key)
                if prev is None:
                    winners[dedup_key] = cand
                else:
                    better = False
                    if source_date != prev["source_date"]:
                        better = source_date > prev["source_date"]
                    elif int(cand["doc_priority"]) != int(prev["doc_priority"]):
                        better = int(cand["doc_priority"]) > int(prev["doc_priority"])
                    elif abs(float(cand["score"]) - float(prev["score"])) > 1e-9:
                        better = float(cand["score"]) > float(prev["score"])
                    if better:
                        winners[dedup_key] = cand

    def _fallback_promise_records() -> List[Dict[str, Any]]:
        out: List[Dict[str, Any]] = []
        deny_re = re.compile(
            r"\b(no assurance|forward-looking statements|private securities litigation reform act|safe harbor|"
            r"holders?\s+of\s+the\s+.*notes?|indenture|fundamental change|hedged transactions|"
            r"forecasted to occur|for the (?:three|nine) months ended|compared with the same period|"
            r"results? for the (?:first|second|third|fourth) quarter)\b",
            re.I,
        )
        preferred_doc_re = re.compile(
            r"(pressrelease|earningsrelease|presentation|slides|transcript|exhibit99|ex-99|exhibit991|shareholder)",
            re.I,
        )
        for _, r in p.iterrows():
            pid = str(r.get(pid_col) or "").strip()
            if not pid:
                continue
            raw_txt = str(r.get(txt_col) or "")
            txt = glx_normalize_text(re.sub(r"\s*created\s+\d{4}-\d{2}-\d{2}\s*$", "", raw_txt, flags=re.I))
            if not txt or deny_re.search(txt) or market_backdrop_re.search(txt):
                continue
            created_q = _qend(r.get(created_col) if created_col else None)
            first_seen_ev_q = _qend(r.get(first_seen_ev_col) if first_seen_ev_col else None) or created_q
            last_seen_ev_q = _qend(r.get(last_seen_ev_col) if last_seen_ev_col else None) or _qend(r.get(last_seen_col) if last_seen_col else None) or first_seen_ev_q
            display_q = last_seen_ev_q or first_seen_ev_q or created_q
            if display_q is None:
                continue
            if q_set and display_q not in q_set:
                continue
            src = _source_meta(r)
            src_text = " ".join(
                [
                    str(src.get("source_type") or ""),
                    str(src.get("doc") or ""),
                    str(src.get("form") or ""),
                ]
            )
            preferred_src = bool(preferred_doc_re.search(src_text))
            score_info = qn_score_promise_candidate(
                text=txt,
                source_type=str(src.get("source_type") or ""),
                form=str(src.get("form") or ""),
                doc_name=str(src.get("doc") or ""),
                section=str(src.get("section") or ""),
            )
            if bool(score_info.get("hard_exclude")):
                continue
            metric_txt = _clean_metric(r.get(metric_col) if metric_col else "")
            metric_canon = _map_raw_metric(metric_txt)
            if re.search(r"\b(annualized cost reductions?|additional annualized savings|expense reduction initiative|cost reduction initiative)\b", txt, re.I):
                metric_canon = "Cost savings"
            elif re.search(r"\b(one-time charge|approximately\s+\$[0-9].{0,40}\$[0-9])\b", txt, re.I):
                metric_canon = "Restructuring charges"
            elif re.search(r"\b(45z|clean fuel production credits?|tax credits?)\b", txt, re.I):
                metric_canon = "Capital allocation"
            if metric_canon in {"", "Milestone"} and not (
                preferred_src
                and bool(re.search(r"\b(online|fully operational|ramping|commissioning|completed|on track)\b", txt, re.I))
                and bool(re.search(r"\b(202\d|q[1-4]|full[- ]?year|by (?:the end of )?20\d{2}|next year|next quarter)\b", txt, re.I))
            ):
                continue
            if metric_canon in {"", "Milestone"}:
                metric_canon = "Strategic milestone"
            has_numeric_target = _has_structured_numeric_target(txt, require_money_like=False)
            if historical_results_re.search(txt):
                continue
            if not (
                re.search(r"\b(target|targets|targeting|expect|expects|expected|plan|plans|will|annualized|agreement|sale agreement|benefit)\b", txt, re.I)
                or has_numeric_target
            ):
                continue
            if metric_canon == "Strategic milestone":
                if not preferred_src:
                    continue
            elif metric_canon not in {"Cost savings", "Restructuring charges"} and not has_numeric_target:
                continue
            if metric_canon in {"Cost savings", "Restructuring charges"} and not (has_numeric_target or fallback_allow_numeric_re.search(txt)):
                continue
            if not preferred_src and float(score_info.get("score") or 0.0) < 24.0:
                continue
            period_key = str(score_info.get("period_key") or "").strip()
            period_label = str(score_info.get("period_label") or "").strip()
            if metric_canon == "Cost savings" and not period_key:
                period_key = "ANNUALIZED_PROGRAM"
                period_label = "Annualized program"
            if metric_canon == "Restructuring charges" and not period_key:
                period_key = "TIME_ANCHOR"
                period_label = period_label or "Time anchor"
            if metric_canon == "Capital allocation" and re.search(r"\b(45z|tax credits?)\b", txt, re.I):
                period_key = period_key or "PROGRAM"
                period_label = period_label or "Program / agreement"
            if not period_key and not preferred_src:
                continue
            rec = {
                "promise_id": pid,
                "metric": metric_canon,
                "period_label": period_label,
                "period_key": period_key or "FALLBACK",
                "quarter": display_q,
                "text_full": txt,
                "text_snippet": qn_compact_snippet(txt, 240),
                "score": max(34.0, float(score_info.get("score") or 0.0) + (4.0 if preferred_src else 0.0)),
                "source": src,
                "source_date": _to_ts(src.get("filed") or r.get(last_seen_col) or r.get(created_col)),
                "doc_priority": int(score_info.get("doc_priority") or 0) + (2 if preferred_src else 0),
                "reasons": ["fallback_recall"],
                "has_numeric": _has_structured_numeric_target(txt, require_money_like=False),
                "has_time_anchor": bool(score_info.get("has_time_anchor")) or bool(period_key),
                "guidance_type": _promise_guidance_type(metric_canon, txt, period_key),
                "as_of_quarter_end": str(display_q),
                "source_doc_end": str(display_q),
                "source_filed_date": pd.to_datetime(src.get("filed"), errors="coerce"),
                "first_seen_quarter_end": str(first_seen_ev_q) if first_seen_ev_q is not None else str(display_q),
                "last_seen_quarter_end": str(last_seen_ev_q) if last_seen_ev_q is not None else str(display_q),
                "referenced_years": sorted({int(y) for y in re.findall(r"(?<!\d)(20\d{2})(?!\d)", txt)}),
                "has_forward_intent": True,
                "has_period_anchor": bool(period_key),
                "target_period_norm": period_key,
                "promise_type": "milestone" if metric_canon == "Strategic milestone" else "operational",
            }
            rec.update(_derive_split_target_meta(metric_canon, txt, period_key, display_q, src.get("source_type"), src.get("doc"), src.get("section")))
            rec["promise_group"] = str(rec.get("target_group_key") or "")
            rec["metric_display"] = _split_target_metric_display(metric_canon, txt, rec)
            out.append(rec)
        dedup: Dict[Tuple[date, str, str], Dict[str, Any]] = {}
        for cand in out:
            k = (cand["quarter"],) + _split_target_identity_key(cand, cand.get("metric"), cand.get("period_key"), cand.get("quarter"))
            prev = dedup.get(k)
            if prev is None or (
                float(cand.get("score") or 0.0),
                int(cand.get("doc_priority") or 0),
            ) > (
                float(prev.get("score") or 0.0),
                int(prev.get("doc_priority") or 0),
            ):
                dedup[k] = cand
        return list(dedup.values())

    fallback_candidates = _fallback_promise_records()
    if not winners:
        for cand in fallback_candidates:
            winners[_split_target_identity_key(cand, cand.get("metric"), cand.get("period_key"), cand.get("quarter"))] = cand
        if winners:
            ui_info_rows.append(
                {
                    "quarter": None,
                    "metric": "Promise_Tracker_UI",
                    "severity": "info",
                    "message": f"fallback_recall_from_raw count={len(winners)}",
                    "source": "",
                }
            )
        else:
            if not has_qnote_recall_seed:
                _store_tracker_state({}, [])
                if render_visible and ws is not None:
                    ws["A1"] = f"Generated at {ts} | Quarter list view"
                    ws["A2"] = "No promises after high-signal filtering."
                    ws.freeze_panes = "A2"
                return qa_rows
    elif fallback_candidates:
        winners_per_quarter: Dict[date, int] = {}
        for rec in winners.values():
            qd = rec.get("quarter")
            if isinstance(qd, date):
                winners_per_quarter[qd] = winners_per_quarter.get(qd, 0) + 1
        added_sparse_fallback = 0
        for cand in fallback_candidates:
            qd = cand.get("quarter")
            if not isinstance(qd, date):
                continue
            if winners_per_quarter.get(qd, 0) >= 2:
                continue
            dedup_key = _split_target_identity_key(cand, cand.get("metric"), cand.get("period_key"), cand.get("quarter"))
            prev = winners.get(dedup_key)
            if prev is not None and _promise_quality_key(cand) >= _promise_quality_key(prev):
                continue
            winners[dedup_key] = cand
            winners_per_quarter[qd] = winners_per_quarter.get(qd, 0) + 1
            added_sparse_fallback += 1
        if added_sparse_fallback > 0:
            ui_info_rows.append(
                {
                    "quarter": None,
                    "metric": "Promise_Tracker_UI",
                    "severity": "info",
                    "message": f"sparse_quarter_fallback_recall count={added_sparse_fallback}",
                    "source": "",
                }
            )

    def _promise_tracker_keep_item(item: Dict[str, Any]) -> bool:
        metric_name = str(item.get("metric") or "").strip()
        metric_low = metric_name.lower()
        txt_full = glx_normalize_text(str(item.get("text_full") or item.get("text_snippet") or ""))
        txt_raw = glx_normalize_text(str(item.get("text_full_raw") or ""))
        def _progress_metric_label_is_fragment(metric_in: str) -> bool:
            metric_txt = glx_normalize_text(str(metric_in or ""))
            if not metric_txt:
                return True
            return bool(
                re.search(
                    r"^\s*(which time|portion of|at which time|for which|that will|who will|where the|"
                    r"the partnership|the merger|the transactions?)\b",
                    metric_txt,
                    re.I,
                )
            )
        is_45z_metric = bool(re.search(r"\b45z\b|tax credit", metric_low, re.I))
        if _progress_metric_label_is_fragment(metric_name):
            return False
        if _slide_signal_noise(txt_raw or txt_full):
            return False
        src = dict(item.get("source") or {})
        src_type = str(src.get("source_type") or src.get("doc_type") or item.get("source_type") or item.get("doc_type") or "").lower()
        preferred_src = bool(preferred_promise_source_re.search(src_type)) or any(
            k in src_type for k in ("presentation", "release", "transcript", "ceo")
        )
        source_class = shared_source_class(src_type)
        statement_class = shared_statement_class(
            txt_full,
            source_type=src_type,
            metric_hint=" | ".join([metric_name, str(item.get("metric_display") or ""), str(item.get("target_display") or item.get("target") or "")]),
        )
        explicit_timing = bool(re.search(r"\b(fy\s*20\d{2}|20\d{2}|q[1-4]|quarter|full[- ]?year|annualized)\b", txt_full, re.I))
        numeric_target = bool(re.search(r"\$?\s*\d+(?:\.\d+)?\s*(?:m|mm|million|b|bn|%|x)?", txt_full, re.I))
        action_or_target = bool(
            re.search(
                r"\b(target|guidance|expected|expect|opportunity|on track|fully operational|online|ramping|completed|executed|sale completed|repaid|savings)\b",
                txt_full,
                re.I,
            )
        )
        summary_text = glx_normalize_text(str(item.get("statement_summary") or ""))
        qualified_promise = shared_qualify_promise_candidate(
            txt_full,
            source_type=str(src.get("source_type") or src.get("doc_type") or item.get("source_type") or item.get("doc_type") or "promise_tracker_ui"),
            metric_hint=" | ".join([
                metric_name,
                str(item.get("metric_display") or ""),
                str(item.get("target_display") or item.get("target") or ""),
            ]),
        )
        if qualified_promise is None and summary_text and summary_text != txt_full:
            qualified_promise = shared_qualify_promise_candidate(
                summary_text,
                source_type=str(src.get("source_type") or src.get("doc_type") or item.get("source_type") or item.get("doc_type") or "promise_tracker_ui"),
                metric_hint=" | ".join([
                    metric_name,
                    str(item.get("metric_display") or ""),
                    str(item.get("target_display") or item.get("target") or ""),
                ]),
            )
        statement_role, statement_drop_reason = shared_classify_statement_evidence_role(
            summary_text or txt_full,
            source_type=str(src.get("source_type") or src.get("doc_type") or item.get("source_type") or item.get("doc_type") or "promise_tracker_ui"),
            metric_hint=" | ".join([
                metric_name,
                str(item.get("metric_display") or ""),
                str(item.get("target_display") or item.get("target") or ""),
            ]),
            target_period_norm=str(item.get("target_period_norm") or item.get("period_norm") or ""),
            promise_type=str(item.get("promise_type") or ""),
        )
        item["evidence_role"] = str(item.get("evidence_role") or statement_role or "promise_origin")
        if statement_role in {"later_evidence", "result_evidence"}:
            item["drop_reason"] = str(item.get("drop_reason") or "later_evidence_not_tracker_origin")
            return False
        explicit_result_only = bool(
            re.search(
                r"\b(fully operational|fully online|online and ramping|repaid|repayment completed|reduced debt|reducing debt|"
                r"principal debt reduction|reduced principal debt|repurchas\w*|expanded margin|margin expanded|"
                r"operating expenses declined|opex declined|achieved)\b",
                summary_text or txt_full,
                re.I,
            )
        )
        explicit_commitment_anchor = bool(
            re.search(
                r"\b(target|guidance|expect|expected|plan|plans|intend|intends|on track|by end of|deadline|goal|will be|to be)\b",
                summary_text or txt_full,
                re.I,
            )
        )
        if explicit_result_only and not explicit_commitment_anchor:
            item["drop_reason"] = str(item.get("drop_reason") or "later_evidence_not_tracker_origin")
            return False
        if qualified_promise is None:
            gpre_targeted_metric = bool(
                is_gpre_profile
                and (
                    is_45z_metric
                    or re.search(r"\b(debt reduction|cost savings|carbon capture|strategic milestone)\b", metric_low, re.I)
                )
            )
            drop_reason = shared_promise_candidate_drop_reason(
                txt_full,
                source_type=str(src.get("source_type") or src.get("doc_type") or item.get("source_type") or item.get("doc_type") or "promise_tracker_ui"),
                metric_hint=" | ".join([
                    metric_name,
                    str(item.get("metric_display") or ""),
                    str(item.get("target_display") or item.get("target") or ""),
                ]),
            )
            gpre_fallback_ok = bool(
                gpre_targeted_metric
                and preferred_src
                and not drop_reason
                and not is_tabular_fragment(txt_full)
                and (
                    numeric_target
                    or explicit_timing
                    or action_or_target
                    or re.search(
                        r"\b(fully operational|online and ramping|agreement executed|repaid|repayment completed|construction progressing|commissioning|on track)\b",
                        txt_full,
                        re.I,
                    )
                )
            )
            if not gpre_fallback_ok:
                if statement_drop_reason:
                    item["drop_reason"] = str(item.get("drop_reason") or statement_drop_reason)
                return False
            item.setdefault("statement_summary", qn_compact_snippet(txt_full, 180))
            item.setdefault(
                "candidate_scope",
                "milestone"
                if metric_low == "strategic milestone"
                else "hard_target",
            )
        else:
            item.setdefault("statement_summary", qualified_promise.summary)
            item.setdefault("candidate_scope", qualified_promise.scope)
        if source_class in {"weak_support", "support"} and statement_class not in {"structured_numeric_bridge"}:
            return False
        if statement_class in {"boilerplate", "scaffolding", "fragmentary_text", "weak_forward_looking"}:
            return False
        completion_like = bool(re.search(r"\b(fully operational|fully online|agreement executed|repaid|repayment completed|completed)\b", txt_full, re.I))
        future_anchor = bool(re.search(r"\b(target|guidance|expect|expected|on track|will|by\s+20\d{2}|q[1-4]\s*20\d{2}|full[- ]?year)\b", txt_full, re.I))
        if completion_like and not future_anchor and not numeric_target:
            return False
        if is_45z_metric and _is_45z_crush_margin_support_only(txt_full):
            return False
        if metric_low == "revenue_yoy":
            return False
        if metric_low in {"capital_allocation", "management target", "tone | corporate"}:
            return False
        if is_pbi_profile:
            source_type = str(src.get("source_type") or "").lower()
            specific_label = _classify_pbi_metric_label(
                " | ".join([metric_name, txt_full, txt_raw, str(item.get("metric_display") or "")]),
                metric_name,
            )
            target_display = str(item.get("target_display") or "").strip() or _extract_pbi_target_display(txt_full, metric_name)
            clean_sentence = _is_pbi_clean_sentence(txt_full)
            effective_label = specific_label or metric_name
            if effective_label in {"", "Management target", "Operating target"}:
                return False
            if source_type == "guidance_snapshot":
                item["metric_display"] = effective_label
                if target_display:
                    item["target_display"] = target_display
                return effective_label in {
                    "Adjusted EBIT guidance",
                    "Revenue guidance",
                    "EPS guidance",
                    "FCF target",
                } and _pbi_target_display_ok(target_display)
            if _looks_pbi_fragment_text(txt_full):
                return False
            if not _pbi_promise_theme_re.search(f"{metric_name} | {txt_full}"):
                return False
            if not preferred_src:
                return False
            item["metric_display"] = effective_label
            if target_display:
                item["target_display"] = target_display
            if effective_label == "Strategic milestone":
                return clean_sentence and explicit_timing and action_or_target
            if effective_label in {
                "Adjusted EBIT guidance",
                "Revenue guidance",
                "EPS guidance",
                "FCF target",
                "Cost savings target",
                "Cost savings program",
                "Cost savings tranche 1",
                "Cost savings tranche 2",
                "PB Bank liquidity release",
                "Deleveraging target",
                "SendTech / Presort operating target",
            }:
                return _pbi_target_display_ok(target_display) and (explicit_timing or source_type == "guidance_snapshot")
            return clean_sentence and (numeric_target or explicit_timing) and action_or_target
        if is_gpre_profile and (
            is_45z_metric
            or re.search(r"\b(debt reduction|cost savings|carbon capture|strategic milestone)\b", metric_low, re.I)
        ):
            gpre_result_only = bool(
                re.search(
                    r"\b(fully operational|fully online|online and ramping|repaid|repayment completed|sale proceeds used to repay|"
                    r"realized|recognized|repurchas\w*|expanded margin|operating expenses declined)\b",
                    txt_full,
                    re.I,
                )
            )
            gpre_forward_anchor = bool(
                re.search(
                    r"\b(target|guidance|expected|expect|on track|by\s+20\d{2}|q[1-4]\s*20\d{2}|2026|2025|full[- ]?year|opportunity)\b",
                    txt_full,
                    re.I,
                )
            )
            if gpre_result_only and not gpre_forward_anchor and not numeric_target:
                return False
            if metric_low == "strategic milestone" and re.search(
                r"\b(online and ramping(?: up capture volumes)?|fully online delivering biogenic co2)\b",
                txt_full,
                re.I,
            ) and not re.search(r"\b(york|advantage nebraska|central city|wood river|obion)\b", txt_full, re.I):
                return False
            return bool(
                numeric_target
                or explicit_timing
                or gpre_forward_anchor
                or re.search(
                    r"\b(construction progressing|commissioning|on track)\b",
                    txt_full,
                    re.I,
                )
            )
        if metric_low == "utilization":
            return preferred_src and explicit_timing and bool(
                re.search(r"\b(utilization|operating rate|capacity utilization)\b", txt_full, re.I)
                and re.search(r"(?<!\d)\d{2,3}%", txt_full)
                and re.search(r"\b(target|expected|on track|objective|goal|maintain|continue to|will)\b", txt_full, re.I)
            )
        if metric_low == "risk management":
            return preferred_src and explicit_timing and bool(
                re.search(r"\brisk management\b", txt_full, re.I)
                and re.search(r"\b(margins?|cash flow|lock in favorable|protect downside|economics)\b", txt_full, re.I)
                and re.search(r"\b(expected|on track|objective|goal|continue to|will|supports?)\b", txt_full, re.I)
            )
        if metric_low == "strategic milestone":
            return explicit_timing or action_or_target
        if metric_low in {"45z monetization / ebitda", "cost savings", "debt reduction"} or is_45z_metric:
            return True
        if metric_low == "management target":
            return preferred_src and action_or_target and (
                explicit_timing
                or numeric_target
                or bool(re.search(r"\b(annualized|qualify|on track|opportunity|expected)\b", txt_full, re.I))
            )
        return preferred_src and (numeric_target or explicit_timing) and action_or_target
    _pbi_tracker_allowed_labels = {
        "Adjusted EBIT guidance",
        "Revenue guidance",
        "EPS guidance",
        "FCF target",
        "Cost savings target",
        "Cost savings program",
        "Cost savings tranche 1",
        "Cost savings tranche 2",
        "Deleveraging target",
        "PB Bank liquidity release",
        "SendTech / Presort operating target",
        "Strategic milestone",
    }

    def _pbi_tracker_label_alignment_ok(label: str, text_in: Any) -> bool:
        label_txt = str(label or "").strip()
        if not label_txt:
            return False
        txt_low = glx_normalize_text(str(text_in or "")).lower()
        if not txt_low:
            return False
        patterns = {
            "Adjusted EBIT guidance": r"\b(adjusted ebit|adj\.?\s*ebit|ebit|margin|profitabilit)\b",
            "Adjusted EBIT / margin": r"\b(adjusted ebit|adj\.?\s*ebit|ebit|margin|profitabilit)\b",
            "Revenue guidance": r"\b(revenue|sales|volume|mail|shipping)\b",
            "Revenue / volume": r"\b(revenue|sales|volume|mail|shipping)\b",
            "EPS guidance": r"\b(eps|earnings per share)\b",
            "FCF improvement": r"\b(fcf|free cash flow|cash flow)\b",
            "FCF target": r"\b(fcf|free cash flow|cash flow)\b",
            "Cost savings / rationalization": r"\b(cost savings|cost reduction|annualized savings|run-rate|rationalization)\b",
            "Cost savings target": r"\b(cost savings|cost reduction|annualized savings|run-rate|rationalization)\b",
            "Deleveraging / liquidity": r"\b(deleverag|debt|leverage|liquidity|repay|repayment|paydown)\b",
            "Deleveraging target": r"\b(deleverag|debt|leverage|liquidity|repay|repayment|paydown)\b",
            "Debt reduction": r"\b(deleverag|debt|leverage|liquidity|repay|repayment|paydown)\b",
            "PB Bank liquidity release": r"\b(pb bank|bank-held leases|leases held|cash optimization|cash needs reduction|receivables purchase|liquidity|cash release|trapped capital)\b",
            "SendTech / Presort operating driver": r"\b(sendtech|presort)\b",
            "SendTech / Presort operating target": r"\b(sendtech|presort)\b",
        }
        pat = patterns.get(label_txt)
        if not pat:
            return True
        return bool(re.search(pat, txt_low, re.I))

    def _pbi_final_tracker_keep_item(item: Dict[str, Any]) -> bool:
        if not is_pbi_profile:
            return True
        src = dict(item.get("source") or {})
        src_type = str(src.get("source_type") or "").lower()
        txt_full = glx_normalize_text(str(item.get("text_full") or item.get("text_snippet") or ""))
        metric_label = _classify_pbi_metric_label(
            " | ".join(
                [
                    str(item.get("metric_display") or ""),
                    str(item.get("metric") or ""),
                    txt_full,
                    str(item.get("target_display") or ""),
                ]
            ),
            str(item.get("metric_display") or item.get("metric") or ""),
        )
        target_display = str(item.get("target_display") or "").strip() or _extract_pbi_target_display(txt_full, metric_label)
        if metric_label not in _pbi_tracker_allowed_labels:
            return False
        if not _pbi_tracker_label_alignment_ok(metric_label, " | ".join([txt_full, target_display])):
            return False
        if metric_label != "Strategic milestone" and not _pbi_target_display_ok(target_display):
            return False
        if src_type in {
            "guidance_snapshot",
            "pbi_guidance_structured",
            "pbi_promise_structured",
            "pbi_quarter_notes_structured",
        }:
            item["metric_display"] = metric_label
            if target_display:
                item["target_display"] = target_display
            return True
        if metric_label == "Strategic milestone":
            if target_display.lower() in {"", "milestone"} and not _is_pbi_clean_sentence(txt_full):
                return False
            if not re.search(r"\b(20\d{2}|q[1-4]|completed|launched|migrated|exited|shut(?:down)?|sold|sale|separated|wind-?down)\b", txt_full, re.I):
                return False
        if not _is_preferred_narrative_source(src_type):
            return False
        if _looks_pbi_fragment_text(txt_full) or not _is_pbi_clean_sentence(txt_full):
            return False
        item["metric_display"] = metric_label
        if target_display:
            item["target_display"] = target_display
        return True

    if render_visible:
        records = [rec for rec in winners.values() if _promise_tracker_keep_item(rec)]
    else:
        # The visible tracker sheet applies stricter presentation rules than the
        # downstream progress sheet needs for its feeder state. When the tracker
        # UI is intentionally hidden, keep the normalized winners available so
        # Promise_Progress_UI can make the final investor-facing quality choice.
        records = [dict(rec) for rec in winners.values() if str(rec.get("promise_id") or "").strip()]
    grouped: Dict[date, List[Dict[str, Any]]] = {}
    for rec in records:
        grouped.setdefault(rec["quarter"], []).append(rec)
    def _gpre_final_tracker_keep_item(item: Dict[str, Any]) -> bool:
        txt_full = glx_normalize_text(str(item.get("text_full") or item.get("text_snippet") or ""))
        metric_name = _gpre_clean_visible_promise_metric(
            str(item.get("metric_display") or item.get("metric") or "").strip(),
            " | ".join(
                [
                    txt_full,
                    str(item.get("target_display") or ""),
                    str(item.get("latest_display") or ""),
                ]
            ),
            item,
        )
        if metric_name:
            item["metric_display"] = metric_name
        if _gpre_bad_visible_promise_reason(metric_name, txt_full, item.get("latest_display"), item.get("target_display")):
            return False
        src = dict(item.get("source") or {})
        statement_role, _ = shared_classify_statement_evidence_role(
            txt_full,
            source_type=str(src.get("source_type") or src.get("doc_type") or item.get("source_type") or ""),
            metric_hint=" | ".join([metric_name, str(item.get("target_display") or ""), str(item.get("period_label") or "")]),
            target_period_norm=str(item.get("target_period_norm") or item.get("period_key") or ""),
            promise_type=str(item.get("promise_type") or ""),
        )
        if statement_role in {"later_evidence", "result_evidence"}:
            return False
        explicit_timing = bool(re.search(r"\b(fy\s*20\d{2}|20\d{2}|q[1-4]|quarter|full[- ]?year|annualized)\b", txt_full, re.I))
        numeric_target = bool(re.search(r"\$?\s*\d+(?:\.\d+)?\s*(?:m|mm|million|b|bn|%|x)?", txt_full, re.I))
        if re.search(
            r"\b(fully operational|fully online|online and ramping|repaid|repayment completed|sale proceeds used to repay|realized|recognized|repurchas\w*|expanded margin|operating expenses declined)\b",
            txt_full,
            re.I,
        ) and not (explicit_timing and re.search(r"\b(target|guidance|expected|on track|opportunity)\b", txt_full, re.I)):
            return False
        if re.search(r"\b(executed construction management agreements|ordered major equipment|class vi well permit)\b", txt_full, re.I) and not (numeric_target or explicit_timing):
            return False
        if (
            "45z" in metric_name.lower()
            and re.search(r"\b(fully operational|fully online|online and ramping|agreement executed)\b", txt_full, re.I)
            and not re.search(r"\b(target|guidance|expected|on track|q4|2026|\$|\b15\b|\b25\b|monetization value|ebitda)\b", txt_full, re.I)
        ):
            return False
        return True

    for qd in list(grouped.keys()):
        grouped[qd] = sorted(
            grouped[qd],
            key=lambda x: (
                metric_priority.get(str(x.get("metric") or FORWARD_NOTES_LABEL), 99),
                _promise_quality_key(x),
                -float(x.get("score") or 0.0),
                -int(x.get("doc_priority") or 0),
            ),
        )[:15]

    quarter_note_rows_map = ui_state.get("quarter_notes_ui_rows", {}) if isinstance(ui_state, dict) else {}

    def _build_pbi_qnote_tracker_fallback_rows(qd: date) -> List[Dict[str, Any]]:
        if not is_pbi_profile or not isinstance(quarter_note_rows_map, dict):
            return []
        out: List[Dict[str, Any]] = []
        seen_keys: set[Tuple[str, str]] = set()
        for item in quarter_note_rows_map.get(qd, []) or []:
            txt_full = glx_normalize_text(str(item.get("text_full") or item.get("comment_full_text") or ""))
            compact_note = str(item.get("_pbi_compact_note") or item.get("text_snippet") or "").strip()
            compact_clean = glx_normalize_text(compact_note)
            visible_note = compact_clean or qn_compact_snippet(txt_full, 220)
            if not visible_note and not txt_full:
                continue
            if visible_note and _looks_pbi_fragment_text(visible_note):
                visible_note = ""
            if txt_full and _looks_pbi_fragment_text(txt_full) and not visible_note:
                continue
            metric_hint = str(item.get("_metric_display") or item.get("metric_canon") or item.get("metric_tag") or "").strip()
            metric_hint_clean = str(metric_hint.split("|", 1)[0] or "").strip()
            metric_label = _classify_pbi_metric_label(
                " | ".join(
                    [
                        metric_hint,
                        str(item.get("metric_canon") or ""),
                        str(item.get("metric_tag") or ""),
                        visible_note,
                        txt_full,
                    ]
                ),
                metric_hint_clean or str(item.get("metric_canon") or item.get("metric_tag") or ""),
            )
            if metric_label not in _pbi_tracker_allowed_labels:
                continue
            target_display = (
                str(item.get("target_display") or "").strip()
                or _extract_pbi_target_display(visible_note, metric_label)
                or _extract_pbi_target_display(txt_full, metric_label)
            )
            if metric_label != "Strategic milestone" and not _pbi_target_display_ok(target_display):
                continue
            if metric_label == "Strategic milestone" and not re.search(
                r"\b(20\d{2}|q[1-4]|completed|launched|migrated|exited|shut(?:down)?|sold|sale|separated|wind-?down)\b",
                txt_full,
                re.I,
            ):
                continue
            key = (
                metric_label.lower(),
                str(target_display or txt_full[:120]).strip().lower(),
            )
            if key in seen_keys:
                continue
            seen_keys.add(key)
            period_norm = str(item.get("target_period_norm") or item.get("period_key") or item.get("period_norm") or "").strip()
            period_label = str(item.get("period_label") or "").strip()
            parsed_targets = _extract_pbi_guidance_targets_multi(
                " | ".join([visible_note, txt_full]).strip(" |"),
                metric_label,
                qd,
            )
            if parsed_targets:
                for parsed in parsed_targets:
                    if str(parsed.get("label") or "").strip() != metric_label:
                        continue
                    period_norm = str(parsed.get("period_norm") or "").strip()
                    period_label = str(parsed.get("period_label") or "").strip() or _guidance_period_label_from_norm(period_norm, qd)
                    if not target_display:
                        target_display = str(parsed.get("target") or "").strip()
                    break
            src = dict(item.get("source") or {})
            tracker_text = visible_note or txt_full
            if not tracker_text:
                continue
            if metric_label == "Strategic milestone" and not _is_pbi_clean_sentence(tracker_text):
                tracker_text = qn_compact_snippet(txt_full, 220) or tracker_text
            rec = {
                "promise_id": f"pbi_qnote_tracker:{qd.isoformat()}:{metric_label}:{target_display or tracker_text[:48]}",
                "quarter": qd,
                "metric": metric_label,
                "metric_display": metric_label,
                "period_label": period_label,
                "period_key": period_norm,
                "target_period_norm": period_norm,
                "target_display": target_display,
                "text_full": tracker_text,
                "text_snippet": qn_compact_snippet(tracker_text, 220),
                "score": float(item.get("score") or 80.0),
                "doc_priority": 7,
                "source": {
                    "source_type": "pbi_quarter_notes_structured",
                    "doc": str(src.get("doc") or "Quarter_Notes_UI"),
                    "form": str(src.get("form") or ""),
                    "section": str(src.get("section") or ""),
                },
                "guidance_type": "milestone" if metric_label == "Strategic milestone" else "period",
                "has_forward_intent": True,
                "has_time_anchor": True,
                "promise_type": "milestone" if metric_label == "Strategic milestone" else "guidance_range",
                "theme_key": _promise_theme_key(metric_label, txt_full, period_norm or period_label),
            }
            rec.update(
                _derive_split_target_meta(
                    metric_label,
                    " | ".join([target_display, tracker_text]),
                    period_norm or period_label,
                    qd,
                    rec["source"]["source_type"],
                    rec["source"]["doc"],
                    rec["source"]["section"],
                )
            )
            rec["promise_group"] = str(rec.get("target_group_key") or "")
            out.append(rec)
        return out

    def _tracker_metric_from_event(note_item: Dict[str, Any]) -> str:
        event_type = str(note_item.get("_event_type") or "").strip().lower()
        metric_family = str(note_item.get("_event_metric_family") or "").strip().lower()
        entity_scope = str(note_item.get("_event_entity_scope") or "").strip().lower()
        if not event_type and not metric_family:
            return ""
        if is_pbi_profile:
            if event_type == "guidance":
                return {
                    "revenue": "Revenue guidance",
                    "adj_ebit": "Adjusted EBIT guidance",
                    "eps": "EPS guidance",
                    "fcf": "FCF target",
                    "cost_savings": "Cost savings target",
                    "liquidity": "PB Bank liquidity release",
                    "debt": "Deleveraging target",
                }.get(metric_family, "")
            if event_type == "cost_savings" or metric_family == "cost_savings":
                return "Cost savings target"
            if event_type == "liquidity_release" or metric_family == "liquidity" or entity_scope == "pb_bank":
                return "PB Bank liquidity release"
            if event_type == "deleveraging" or metric_family == "debt":
                return "Deleveraging target"
            if event_type == "milestone" or metric_family == "milestone":
                return "Strategic milestone"
            return ""
        if event_type == "regulatory_credit" or metric_family == "regulatory_credit":
            return "45Z monetization / EBITDA"
        if event_type == "cost_savings" or metric_family == "cost_savings":
            return "Cost savings"
        if event_type == "deleveraging" or metric_family == "debt" or entity_scope == "obion":
            return "Debt reduction"
        if event_type == "milestone" or metric_family == "milestone":
            return "Strategic milestone"
        return ""

    def _tracker_metric_from_qnote(note_item: Dict[str, Any]) -> str:
        txt_local = glx_normalize_text(str(note_item.get("text_full") or ""))
        hint = str(note_item.get("metric_canon") or note_item.get("metric_tag") or "").strip().lower()
        candidate_type = str(note_item.get("candidate_type") or "").strip().lower()
        event_metric = _tracker_metric_from_event(note_item)
        if event_metric:
            return event_metric
        blob = f"{hint} {txt_local.lower()}"
        if is_pbi_profile:
            pbi_allowed_labels_local = {
                "Adjusted EBIT guidance",
                "Revenue guidance",
                "EPS guidance",
                "FCF target",
                "Cost savings target",
                "PB Bank liquidity release",
                "Deleveraging target",
                "SendTech / Presort operating target",
                "Strategic milestone",
            }
            pbi_metric = _classify_pbi_metric_label(blob, "")
            if pbi_metric in pbi_allowed_labels_local:
                return pbi_metric
            if pbi_metric:
                return ""
        if any(k in blob for k in ("45z", "tax credit monetization", "ebitda opportunity", "qualify for production tax credits")):
            return "45Z monetization / EBITDA"
        if re.search(r"\b(cost reduction|cost savings|annualized savings|expense reduction)\b", blob, re.I):
            return "Cost savings"
        if re.search(r"\b(repay|repaid|delever|debt reduction|used to fully repay|sale of obion)\b", blob, re.I):
            return "Debt reduction"
        if re.search(
            r"\b(fully operational|online|ramping|progressing|under construction|construction progressing|"
            r"start-?up|started up|delivered|received .*permit|permit|commissioning|executed|ordered major equipment|"
            r"construction management agreements?)\b",
            blob,
            re.I,
        ):
            return "Strategic milestone"
        if candidate_type in {"program_line", "slides_priority_signal"} or re.search(
            r"\b(target|expected|opportunity|on track|qualify|will|plan|plans|intend)\b",
            txt_local,
            re.I,
        ):
            return "Management target"
        return ""

    def _tracker_target_display_from_qnote(qd: date, metric_name: str, text_in: Any) -> str:
        txt_local = glx_normalize_text(str(text_in or ""))
        metric_txt = str(metric_name or "").strip()
        if not txt_local or not metric_txt:
            return ""
        if is_pbi_profile:
            return _extract_pbi_target_display(txt_local, metric_txt)
        metric_low = metric_txt.lower()
        if re.search(r"\b45z\b|tax credit", metric_low, re.I):
            return (
                _extract_45z_monetization_target_display(txt_local, qd)
                or _strong_45z_2026_target_display(txt_local, qd, "")
                or ""
            )
        if re.search(r"\b(cost savings|cost reduction|expense reduction)\b", metric_low, re.I):
            amounts = _extract_money_targets_for_display(txt_local)
            if len(amounts) >= 2:
                lo = min(float(amounts[0]), float(amounts[1]))
                hi = max(float(amounts[0]), float(amounts[1]))
                return f"{_fmt_short_money_value_local(lo)}-{_fmt_short_money_value_local(hi)}"
            if amounts:
                return f">= {_fmt_short_money_value_local(float(max(amounts)))}"
        if re.search(r"\bdebt reduction\b", metric_low, re.I):
            amounts = _extract_money_targets_for_display(txt_local)
            if amounts:
                return _fmt_short_money_value_local(float(max(amounts)))
        return ""

    def _tracker_candidate_from_qnote(qd: date, note_item: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        txt_full = glx_normalize_text(str(note_item.get("text_full") or ""))
        if not txt_full or _slide_signal_noise(txt_full):
            return None
        metric_name = _tracker_metric_from_qnote(note_item)
        if not metric_name:
            return None
        source = dict(note_item.get("source") or {})
        metric_priority_bonus = 3 if metric_name in {"45Z monetization / EBITDA", "Cost savings", "Debt reduction"} else 0
        period_label = ""
        m_year = re.search(r"\b(20\d{2})\b", txt_full)
        if m_year:
            period_label = f"FY {m_year.group(1)}"
        summary_txt = glx_normalize_text(str(note_item.get("_render_summary") or note_item.get("statement_summary") or ""))
        if not summary_txt:
            summary_txt = qn_compact_snippet(txt_full, 220)
        rec = {
            "promise_id": str(note_item.get("note_id") or hashlib.sha1(f"{qd}|qnote_tracker|{txt_full}".encode("utf-8")).hexdigest()[:12]),
            "quarter": qd,
            "metric": metric_name,
            "text_snippet": summary_txt,
            "text_full": txt_full,
            "statement_summary": summary_txt,
            "period_label": period_label,
            "guidance_type": "milestone" if metric_name == "Strategic milestone" else "period",
            "first_seen_quarter_end": str(qd),
            "last_seen_quarter_end": str(qd),
            "source": {
                "source_type": str(
                    source.get("source_type")
                    or source.get("doc_type")
                    or note_item.get("source_type")
                    or note_item.get("doc_type")
                    or note_item.get("source_doc_type")
                    or "quarter_notes_ui"
                ),
                "doc": str(source.get("doc") or note_item.get("doc") or "Quarter_Notes_UI"),
                "form": str(source.get("form") or note_item.get("form") or ""),
                "section": str(source.get("section") or note_item.get("section") or ""),
            },
            "score": float(note_item.get("score") or 0.0) + 2.0 + float(metric_priority_bonus),
            "doc_priority": 7,
            "reasons": ["quarter_notes_ui_recall"],
            "theme_key": str(note_item.get("theme_key") or note_item.get("_event_key") or _promise_theme_key(metric_name, txt_full, period_label)),
            "_fragment_penalty": _text_fragment_penalty(txt_full),
            "_clean_target_bonus": _clean_target_bonus(txt_full),
            "target_display": _tracker_target_display_from_qnote(qd, metric_name, txt_full),
            "_event_key": str(note_item.get("_event_key") or ""),
            "_event_type": str(note_item.get("_event_type") or ""),
            "_event_metric_family": str(note_item.get("_event_metric_family") or ""),
            "_event_entity_scope": str(note_item.get("_event_entity_scope") or ""),
            "_event_period_norm": str(note_item.get("_event_period_norm") or note_item.get("target_period_norm") or note_item.get("period_key") or ""),
        }
        if metric_name == "Debt reduction":
            if re.search(r"\b(fully repay|fully repaid|repaid|repayment completed)\b", txt_full, re.I):
                rec["status_hint"] = "completed"
                rec["latest_display"] = "Debt repaid"
            elif re.search(r"\b(deleverag|paydown|repayment)\b", txt_full, re.I):
                rec["status_hint"] = "in progress"
                rec["latest_display"] = "Debt reduction underway"
        elif metric_name == "Strategic milestone":
            if re.search(r"\bfully operational\b", txt_full, re.I):
                rec["status_hint"] = "completed"
                rec["latest_display"] = "Fully operational"
            elif re.search(r"\bonline and ramping\b", txt_full, re.I):
                rec["status_hint"] = "in progress"
                rec["latest_display"] = "Online and ramping"
            elif re.search(r"\bagreement executed\b", txt_full, re.I):
                rec["status_hint"] = "completed"
                rec["latest_display"] = "Agreement executed"
        rec.update(
            _derive_split_target_meta(
                metric_name,
                txt_full,
                note_item.get("target_period_norm") or note_item.get("period_key") or period_label,
                qd,
                source.get("source_type") or "quarter_notes_ui",
                source.get("doc") or "Quarter_Notes_UI",
                source.get("section") or "",
            )
        )
        rec["promise_group"] = str(rec.get("target_group_key") or "")
        rec["metric_display"] = _split_target_metric_display(metric_name, txt_full, rec)
        return rec

    def _tracker_identity_key_local(item: Dict[str, Any]) -> Tuple[str, str, str]:
        canonical_key = str(item.get("canonical_subject_key") or "").strip()
        lifecycle_key = str(item.get("promise_lifecycle_key") or item.get("lifecycle_key") or "").strip()
        lifecycle_subject_key = str(item.get("lifecycle_subject_key") or "").strip()
        if not canonical_key:
            period_norm = shared_infer_target_period_norm(
                period_norm=item.get("target_period_norm") or item.get("period_key") or item.get("period_label") or "",
                deadline=item.get("deadline") or item.get("target_time"),
                quarter=item.get("quarter"),
                text=" | ".join(
                    [
                        str(item.get("target_display") or item.get("target") or ""),
                        str(item.get("text_full") or item.get("text_snippet") or ""),
                        str(item.get("metric_display") or item.get("metric") or ""),
                    ]
                ),
            )
            routed = shared_route_to_measurable_promise_candidate(
                " | ".join(
                    [
                        str(item.get("target_display") or item.get("target") or ""),
                        str(item.get("text_full") or item.get("text_snippet") or ""),
                        str(item.get("latest_display") or item.get("latest") or ""),
                    ]
                ),
                quarter=item.get("quarter"),
                source_type=str(dict(item.get("source") or {}).get("source_type") or item.get("source_type") or "promise_tracker_ui"),
                metric_hint=str(item.get("metric_display") or item.get("metric") or ""),
                source_doc=str(dict(item.get("source") or {}).get("doc") or item.get("doc") or ""),
                target_period_norm=period_norm,
                promise_type_hint=str(item.get("promise_type") or ""),
                base_score=float(item.get("score") or item.get("_score") or 0.0),
            )
            if routed is not None:
                canonical_key = routed.canonical_subject_key
                lifecycle_key = routed.lifecycle_key
                lifecycle_subject_key = routed.lifecycle_subject_key or routed.lifecycle_key
                item["canonical_subject_key"] = canonical_key
                item["promise_lifecycle_key"] = lifecycle_key
                item["lifecycle_subject_key"] = lifecycle_subject_key
                item["parent_subject_key"] = routed.parent_subject_key
                item["routing_reason"] = routed.routing_reason
                item["source_class"] = routed.source_class
                item["statement_class"] = routed.statement_class
                item["evidence_role"] = routed.evidence_role
                item["metric_family"] = routed.metric_family
                item["entity_scope"] = routed.entity_scope
                item["target_period_norm"] = routed.target_period_norm or period_norm
            else:
                parent_subject_key = shared_build_parent_subject_key(
                    entity_scope=item.get("scope_key") or "company_total",
                    metric_family=item.get("metric_display") or item.get("metric") or "general",
                    program_token=item.get("scope_key") or "",
                    topic_family=item.get("bucket") or "",
                )
                canonical_key = shared_build_canonical_subject_key(
                    entity_scope=item.get("scope_key") or "company_total",
                    metric_family=item.get("metric_display") or item.get("metric") or "general",
                    target_period_norm=period_norm,
                    scope_token=item.get("target_group_key") or item.get("theme_key") or "",
                )
                lifecycle_key = shared_build_promise_lifecycle_key(
                    canonical_key,
                    stage_token=item.get("promise_type") or "",
                )
                lifecycle_subject_key = shared_build_lifecycle_subject_key(
                    parent_subject_key=parent_subject_key,
                    canonical_subject_key=canonical_key,
                    stage_token=item.get("promise_type") or "",
                    target_period_norm=period_norm,
                )
                item["parent_subject_key"] = parent_subject_key
                item["canonical_subject_key"] = canonical_key
                item["promise_lifecycle_key"] = lifecycle_key
                item["lifecycle_subject_key"] = lifecycle_subject_key
                item["source_class"] = shared_source_class(dict(item.get("source") or {}).get("source_type") or item.get("source_type") or "")
                item["statement_class"] = shared_statement_class(
                    " | ".join([str(item.get("text_full") or item.get("text_snippet") or ""), str(item.get("target_display") or "")]),
                    source_type=dict(item.get("source") or {}).get("source_type") or item.get("source_type") or "",
                    metric_hint=str(item.get("metric_display") or item.get("metric") or ""),
                )
                item["evidence_role"] = shared_evidence_role("measurable_promise_candidate", route_reason=item.get("route_reason") or item.get("routing_reason") or "promise_tracker", promise_type=item.get("promise_type") or "")
        item["candidate_type"] = str(item.get("candidate_type") or "measurable_promise_candidate")
        item["route_reason"] = str(item.get("route_reason") or item.get("routing_reason") or "promise_tracker")
        stated_q = str(item.get("first_seen_quarter_end") or item.get("quarter") or "")
        latest_q = str(item.get("last_seen_quarter_end") or item.get("quarter") or stated_q)
        carried_q = str(item.get("carried_to_quarter_end") or latest_q or stated_q)
        item["stated_quarter"] = stated_q
        item["latest_evidence_quarter"] = latest_q
        item["evaluated_through_quarter"] = str(item.get("evaluated_through_quarter") or carried_q or latest_q or stated_q)
        item["carried_to_quarter"] = str(item.get("carried_to_quarter") or carried_q or latest_q or stated_q)
        item["lifecycle_state"] = str(
            item.get("lifecycle_state")
            or shared_derive_lifecycle_state(
                target_period_norm=item.get("target_period_norm") or item.get("period_key") or item.get("period_label") or "",
                stated_quarter=stated_q,
                latest_evidence_quarter=latest_q,
                evaluated_through_quarter=item.get("evaluated_through_quarter") or carried_q or latest_q or stated_q,
                carried_to_quarter=item.get("carried_to_quarter") or carried_q or latest_q or stated_q,
                current_status=item.get("status_hint") or item.get("status") or "stated",
            )
        )
        item["status_resolution_reason"] = str(
            item.get("status_resolution_reason")
            or shared_derive_status_resolution_reason(
                current_status=item.get("status_hint") or item.get("status") or "",
                latest_value=item.get("latest_display") or item.get("latest") or "",
                lifecycle_state=item.get("lifecycle_state") or "",
            )
        )
        return _split_target_identity_key(
            {
                **item,
                "target_group_key": lifecycle_subject_key or lifecycle_key or canonical_key or item.get("target_group_key") or "",
            },
            item.get("metric_display") or item.get("metric"),
            item.get("target_period_norm") or item.get("period_key") or item.get("period_label"),
            item.get("quarter"),
        )

    def _display_tracker_metric(item: Dict[str, Any]) -> str:
        metric_display = str(item.get("metric_display") or "").strip()
        metric_display_low = metric_display.lower()
        if is_pbi_profile:
            better_pbi = _classify_pbi_metric_label(
                " | ".join(
                    [
                        str(item.get("target_display") or ""),
                        str(item.get("text_full") or item.get("text_snippet") or ""),
                        metric_display,
                    ]
                ),
                metric_display,
            )
            if better_pbi and better_pbi.lower() not in {"management target"}:
                return better_pbi
        if metric_display and metric_display_low not in {
            "management target",
            "strategic milestone",
            "cost savings",
            "production tax 45z generation",
            "tax 45z generation",
            "qualify for production tax 45z generation",
            "fourth quarter 45z generation",
        }:
            return metric_display
        blob = glx_normalize_text(
            " | ".join(
                [
                    str(item.get("target_display") or ""),
                    str(item.get("text_full") or item.get("text_snippet") or ""),
                ]
            )
        )
        better = _split_target_metric_display(
            item.get("metric"),
            blob,
            item,
        )
        return str(better or metric_display or item.get("metric") or "").strip()

    def _guidance_snapshot_items_for_qd(qd: date) -> List[Dict[str, Any]]:
        if not isinstance(gstore, dict) or not gstore:
            return []
        keys: List[Any] = [str(qd), qd.isoformat(), qd]
        try:
            keys.extend([pd.Timestamp(qd), pd.Timestamp(qd).to_pydatetime()])
        except Exception:
            pass
        items: List[Dict[str, Any]] = []
        seen: set[Tuple[str, str]] = set()
        for key in keys:
            rows = gstore.get(key, [])
            if not isinstance(rows, list):
                continue
            for row in rows:
                if not isinstance(row, dict):
                    continue
                row_key = (
                    str(row.get("metric") or "").strip().lower(),
                    str(row.get("target_period_norm") or row.get("period_norm") or "").strip().lower(),
                )
                if row_key in seen:
                    continue
                seen.add(row_key)
                items.append(row)
        return items

    def _build_pbi_tracker_fallback_rows(qd: date) -> List[Dict[str, Any]]:
        if not is_pbi_profile:
            return []
        fallback_rows: List[Dict[str, Any]] = []
        seen_pairs: set[Tuple[str, str]] = set()

        def _append_row(
            metric_label: str,
            target_display: str,
            *,
            text_full: str,
            source_type: str,
            doc: str,
            form: str = "",
            section: str = "",
            period_norm: str = "",
            period_label: str = "",
            score: float = 84.0,
            doc_priority: int = 7,
            latest_display: str = "",
            status_hint: str = "",
        ) -> None:
            metric_label = str(metric_label or "").strip()
            target_display = str(target_display or "").strip()
            text_full = glx_normalize_text(str(text_full or ""))
            if metric_label not in _pbi_tracker_allowed_labels:
                return
            if metric_label != "Strategic milestone" and not _pbi_target_display_ok(target_display):
                return
            pair = (metric_label.lower(), target_display.lower())
            if pair in seen_pairs:
                return
            seen_pairs.add(pair)
            compact_text = text_full or f"{period_label} {metric_label} target {target_display}".strip()
            rec = {
                "promise_id": f"pbi_tracker_fallback:{qd.isoformat()}:{metric_label}:{target_display}",
                "quarter": qd,
                "metric": metric_label,
                "metric_display": metric_label,
                "period_label": period_label,
                "period_key": period_norm,
                "target_period_norm": period_norm,
                "target_display": target_display,
                "text_full": compact_text,
                "text_snippet": qn_compact_snippet(compact_text, 220),
                "score": float(score),
                "doc_priority": int(doc_priority),
                "source": {
                    "source_type": source_type,
                    "doc": doc,
                    "form": form,
                    "section": section,
                },
                "guidance_type": "period",
                "has_forward_intent": True,
                "has_time_anchor": True,
                "promise_type": "guidance_range",
                "theme_key": _promise_theme_key(metric_label, compact_text, period_norm or period_label),
                "latest_display": glx_normalize_text(str(latest_display or "")),
                "status_hint": str(status_hint or "").strip().lower(),
            }
            rec.update(
                _derive_split_target_meta(
                    metric_label,
                    " | ".join([target_display, compact_text]),
                    period_norm or period_label,
                    qd,
                    source_type,
                    doc,
                    section,
                )
            )
            rec["promise_group"] = str(rec.get("target_group_key") or "")
            fallback_rows.append(rec)

        for structured in _pbi_structured_guidance_items_for_qd(qd):
            metric_label = str(structured.get("metric_label") or "").strip()
            target_display = str(structured.get("target_display") or "").strip()
            period_norm = str(structured.get("period_norm") or "").strip()
            period_label = str(structured.get("period_label") or "").strip()
            src = dict(structured.get("source") or {})
            _append_row(
                metric_label,
                target_display,
                text_full=str(structured.get("text_full") or ""),
                source_type=str(src.get("source_type") or "pbi_guidance_structured"),
                doc=str(src.get("doc") or ""),
                form=str(src.get("form") or ""),
                section=str(src.get("section") or ""),
                period_norm=period_norm,
                period_label=period_label,
                score=float(structured.get("score") or 88.0),
                doc_priority=int(structured.get("doc_priority") or 8),
                latest_display=str(structured.get("latest_display") or ""),
                status_hint=str(structured.get("status_hint") or ""),
            )

        for structured in _pbi_structured_strategy_items_for_qd(qd):
            metric_label = str(structured.get("metric_label") or "").strip()
            target_display = str(structured.get("target_display") or "").strip()
            period_norm = str(structured.get("period_norm") or "").strip()
            period_label = str(structured.get("period_label") or "").strip()
            src = dict(structured.get("source") or {})
            _append_row(
                metric_label,
                target_display,
                text_full=str(structured.get("text_full") or ""),
                source_type=str(src.get("source_type") or "pbi_promise_structured"),
                doc=str(src.get("doc") or ""),
                form=str(src.get("form") or ""),
                section=str(src.get("section") or ""),
                period_norm=period_norm,
                period_label=period_label,
                score=float(structured.get("score") or 86.0),
                doc_priority=int(structured.get("doc_priority") or 8),
                latest_display=str(structured.get("latest_display") or ""),
                status_hint=str(structured.get("status_hint") or ""),
            )

        for item in _guidance_snapshot_items_for_qd(qd):
            metric_name = str(item.get("metric") or "").strip()
            metric_label = _classify_pbi_metric_label(metric_name, metric_name)
            target_display = _guidance_value_snip(item)
            period_norm = str(item.get("target_period_norm") or item.get("period_norm") or "").strip()
            period_label = _guidance_period_label_from_norm(period_norm, qd)
            src = dict(item.get("source") or {})
            _append_row(
                metric_label,
                target_display,
                text_full=f"{period_label} {metric_label} target {target_display}".strip(),
                source_type=str(src.get("source_type") or "guidance_snapshot"),
                doc=str(src.get("doc") or ""),
                form=str(src.get("form") or ""),
                section=str(src.get("section") or src.get("section_or_page") or ""),
                period_norm=period_norm,
                period_label=period_label,
                score=float(item.get("score") or 86.0),
                doc_priority=int(item.get("source_priority") or 8),
            )

        if isinstance(slides_guidance, pd.DataFrame) and not slides_guidance.empty:
            sg = slides_guidance.copy()
            if "quarter" in sg.columns:
                sg["quarter_norm"] = pd.to_datetime(sg["quarter"], errors="coerce").dt.date
            else:
                sg["quarter_norm"] = pd.NaT
            try:
                sg_rows = sg.loc[sg["quarter_norm"] == qd].to_dict("records")
            except Exception:
                sg_rows = []
            for rec in sg_rows:
                txt_full = glx_normalize_text(str(rec.get("line") or ""))
                if not txt_full or _slide_signal_noise(txt_full) or _looks_pbi_fragment_text(txt_full):
                    continue
                metric_hint = str(rec.get("metric_hint") or "").strip()
                for parsed in _extract_pbi_guidance_targets_multi(txt_full, metric_hint, qd):
                    metric_label = str(parsed.get("label") or "").strip()
                    target_display = str(parsed.get("target") or "").strip()
                    period_norm = str(parsed.get("period_norm") or "").strip()
                    period_label = str(parsed.get("period_label") or "").strip() or _guidance_period_label_from_norm(period_norm, qd)
                    compact_text = f"{period_label} {metric_label} target {target_display}".strip()
                    _append_row(
                        metric_label,
                        target_display,
                        text_full=compact_text,
                        source_type="pbi_guidance_structured",
                        doc=str(rec.get("doc") or ""),
                        form="presentation",
                        section=f"page {rec.get('page')}" if rec.get("page") not in (None, "") else "",
                        period_norm=period_norm,
                        period_label=period_label,
                        score=84.0,
                        doc_priority=7,
                    )

        if not p.empty:
            for _, rr in p.iterrows():
                display_q = _qend(rr.get(created_col) if created_col else None) or _qend(rr.get(last_seen_col) if last_seen_col else None)
                if display_q != qd:
                    continue
                txt_full = glx_normalize_text(str(rr.get(txt_col) or ""))
                if not txt_full or _looks_pbi_fragment_text(txt_full) or not _is_pbi_clean_sentence(txt_full):
                    continue
                src = _source_meta(rr)
                src_type = str(src.get("source_type") or "")
                if not _is_preferred_narrative_source(src_type):
                    continue
                metric_seed = _map_raw_metric(_clean_metric(rr.get(metric_col) if metric_col else ""))
                metric_label = _classify_pbi_metric_label(
                    " | ".join([metric_seed, txt_full, str(rr.get(target_kind_col) if target_kind_col else "")]),
                    metric_seed,
                )
                if metric_label not in {
                    "Cost savings target",
                    "Deleveraging target",
                    "PB Bank liquidity release",
                    "SendTech / Presort operating target",
                    "Strategic milestone",
                }:
                    continue
                parsed_targets = _extract_pbi_guidance_targets_multi(txt_full, metric_label, qd)
                if parsed_targets:
                    for parsed in parsed_targets:
                        if str(parsed.get("label") or "").strip() != metric_label:
                            continue
                        _append_row(
                            metric_label,
                            str(parsed.get("target") or "").strip(),
                            text_full=txt_full,
                            source_type="pbi_promise_structured",
                            doc=str(src.get("doc") or ""),
                            form=str(src.get("form") or ""),
                            section=str(src.get("section") or ""),
                            period_norm=str(parsed.get("period_norm") or "").strip(),
                            period_label=str(parsed.get("period_label") or "").strip() or _guidance_period_label_from_norm(str(parsed.get("period_norm") or "").strip(), qd),
                            score=83.0,
                            doc_priority=7,
                        )
                    continue
                if metric_label == "Strategic milestone":
                    deadline_q = _qend(rr.get(deadline_col) if deadline_col else None)
                    period_norm = ""
                    period_label = _quarter_lbl(deadline_q) if deadline_q else ""
                    _append_row(
                        metric_label,
                        "",
                        text_full=txt_full,
                        source_type="pbi_promise_structured",
                        doc=str(src.get("doc") or ""),
                        form=str(src.get("form") or ""),
                        section=str(src.get("section") or ""),
                        period_norm=period_norm,
                        period_label=period_label,
                        score=82.0,
                        doc_priority=6,
                    )

        return fallback_rows

    def _build_pbi_tracker_ui_rows(qd: date) -> List[Dict[str, Any]]:
        if not is_pbi_profile:
            return []
        out: List[Dict[str, Any]] = []
        seen_pairs: set[Tuple[str, str, str]] = set()

        def _append_tracker_row(
            metric_label: str,
            target_display: str,
            text_full: str,
            *,
            source_type: str,
            doc: str,
            form: str = "",
            section: str = "",
            period_norm: str = "",
            period_label: str = "",
            score: float = 84.0,
            doc_priority: int = 7,
            latest_display: str = "",
            status_hint: str = "",
        ) -> None:
            metric_label = str(metric_label or "").strip()
            target_display = str(target_display or "").strip()
            text_full = glx_normalize_text(str(text_full or ""))
            if metric_label not in _pbi_tracker_allowed_labels:
                return
            if metric_label != "Strategic milestone" and not _pbi_target_display_ok(target_display):
                return
            if not text_full:
                return
            pair = (
                metric_label.lower(),
                target_display.lower() or text_full[:120].lower(),
                str(period_norm or period_label or "").strip().lower(),
            )
            if pair in seen_pairs:
                return
            seen_pairs.add(pair)
            rec = {
                "promise_id": f"pbi_tracker_ui:{qd.isoformat()}:{metric_label}:{target_display or text_full[:48]}",
                "quarter": qd,
                "metric": metric_label,
                "metric_display": metric_label,
                "period_label": period_label,
                "period_key": period_norm,
                "target_period_norm": period_norm,
                "target_display": target_display,
                "text_full": text_full,
                "text_snippet": qn_compact_snippet(text_full, 220),
                "score": float(score),
                "doc_priority": int(doc_priority),
                "source": {
                    "source_type": source_type,
                    "doc": doc,
                    "form": form,
                    "section": section,
                },
                "guidance_type": "milestone" if metric_label == "Strategic milestone" else "period",
                "has_forward_intent": True,
                "has_time_anchor": True,
                "promise_type": "milestone" if metric_label == "Strategic milestone" else "guidance_range",
                "theme_key": _promise_theme_key(metric_label, text_full, period_norm or period_label),
                "latest_display": glx_normalize_text(str(latest_display or "")),
                "status_hint": str(status_hint or "").strip().lower(),
            }
            rec.update(
                _derive_split_target_meta(
                    metric_label,
                    " | ".join([target_display, text_full]),
                    period_norm or period_label,
                    qd,
                    source_type,
                    doc,
                    section,
                )
            )
            rec["promise_group"] = str(rec.get("target_group_key") or "")
            out.append(rec)

        for structured in _pbi_structured_guidance_items_for_qd(qd):
            metric_label = str(structured.get("metric_label") or "").strip()
            target_display = str(structured.get("target_display") or "").strip()
            src = dict(structured.get("source") or {})
            text_basis = glx_normalize_text(
                str(structured.get("compact_note") or structured.get("text_full") or "")
            )
            _append_tracker_row(
                metric_label,
                target_display,
                text_basis,
                source_type=str(src.get("source_type") or "pbi_guidance_structured"),
                doc=str(src.get("doc") or ""),
                form=str(src.get("form") or ""),
                section=str(src.get("section") or ""),
                period_norm=str(structured.get("period_norm") or "").strip(),
                period_label=str(structured.get("period_label") or "").strip(),
                score=float(structured.get("score") or 88.0),
                doc_priority=int(structured.get("doc_priority") or 8),
                latest_display=str(structured.get("latest_display") or ""),
                status_hint=str(structured.get("status_hint") or ""),
            )

        for structured in _pbi_structured_strategy_items_for_qd(qd):
            metric_label = str(structured.get("metric_label") or "").strip()
            target_display = str(structured.get("target_display") or "").strip()
            src = dict(structured.get("source") or {})
            text_basis = glx_normalize_text(
                str(structured.get("compact_note") or structured.get("text_full") or "")
            )
            _append_tracker_row(
                metric_label,
                target_display,
                text_basis,
                source_type=str(src.get("source_type") or "pbi_promise_structured"),
                doc=str(src.get("doc") or ""),
                form=str(src.get("form") or ""),
                section=str(src.get("section") or ""),
                period_norm=str(structured.get("period_norm") or "").strip(),
                period_label=str(structured.get("period_label") or "").strip(),
                score=float(structured.get("score") or 86.0),
                doc_priority=int(structured.get("doc_priority") or 8),
                latest_display=str(structured.get("latest_display") or ""),
                status_hint=str(structured.get("status_hint") or ""),
            )

        for item in quarter_note_rows_map.get(qd, []) or []:
            raw_metric = str(item.get("_metric_display") or item.get("metric_canon") or item.get("metric_tag") or "").strip()
            note_text = glx_normalize_text(
                str(item.get("_pbi_compact_note") or item.get("text_snippet") or item.get("text_full") or "")
            )
            full_text = glx_normalize_text(str(item.get("text_full") or item.get("comment_full_text") or ""))
            if not note_text and not full_text:
                continue
            metric_label = raw_metric if raw_metric in _pbi_tracker_allowed_labels else _classify_pbi_metric_label(
                " | ".join([raw_metric, note_text, full_text]),
                raw_metric,
            )
            if metric_label not in _pbi_tracker_allowed_labels:
                continue
            target_display = (
                str(item.get("target_display") or "").strip()
                or _extract_pbi_target_display(note_text, metric_label)
                or _extract_pbi_target_display(full_text, metric_label)
            )
            if metric_label != "Strategic milestone" and not _pbi_target_display_ok(target_display):
                continue
            parsed_targets = _extract_pbi_guidance_targets_multi(" | ".join([note_text, full_text]), metric_label, qd)
            period_norm = str(item.get("target_period_norm") or item.get("period_key") or item.get("period_norm") or "").strip()
            period_label = str(item.get("period_label") or "").strip()
            if parsed_targets:
                for parsed in parsed_targets:
                    if str(parsed.get("label") or "").strip() != metric_label:
                        continue
                    period_norm = str(parsed.get("period_norm") or "").strip()
                    period_label = str(parsed.get("period_label") or "").strip() or _guidance_period_label_from_norm(period_norm, qd)
                    if not target_display:
                        target_display = str(parsed.get("target") or "").strip()
                    break
            if metric_label == "Strategic milestone":
                text_basis = note_text or qn_compact_snippet(full_text, 220)
                if not text_basis or not (
                    _is_pbi_clean_sentence(text_basis) or re.search(r"\b(completed|migrated|exited|sold|launched|shut(?:down)?)\b", text_basis, re.I)
                ):
                    continue
            else:
                text_basis = note_text or target_display
            src = dict(item.get("source") or {})
            _append_tracker_row(
                metric_label,
                target_display,
                text_basis,
                source_type="pbi_quarter_notes_structured",
                doc=str(src.get("doc") or "Quarter_Notes_UI"),
                form=str(src.get("form") or ""),
                section=str(src.get("section") or ""),
                period_norm=period_norm,
                period_label=period_label,
                score=float(item.get("score") or 82.0),
                doc_priority=7,
            )

        return out

    def _build_pbi_tracker_rows_from_ui_sheet(qd: date) -> List[Dict[str, Any]]:
        if not is_pbi_profile or "Quarter_Notes_UI" not in wb.sheetnames:
            return []
        out: List[Dict[str, Any]] = []
        seen_pairs: set[Tuple[str, str, str]] = set()
        ws_qn = wb["Quarter_Notes_UI"]
        rows_scanned = 0
        rejected_bad_metric = 0
        rejected_bad_target = 0

        def _ui_note_target(note_in: str) -> str:
            txt_local = glx_normalize_text(str(note_in or ""))
            if not txt_local:
                return ""
            patterns = [
                r"^(?:\[(?:NEW|REPEAT|DROPPED)\]\s*)?(?:updated\s+target|target|raised target to|reaffirmed target|pb bank target|cost savings target)\s+(.+)$",
                r"^(?:\[(?:NEW|REPEAT|DROPPED)\]\s*)?(tracking midpoint of|tracking low end of)\s+(.+)$",
            ]
            for pat in patterns:
                m_direct = re.match(pat, txt_local, re.I)
                if not m_direct:
                    continue
                if m_direct.lastindex and m_direct.lastindex >= 2:
                    return str(m_direct.group(2) or "").strip(" .")
                return str(m_direct.group(1) or "").strip(" .")
            return ""

        def _default_period(metric_label_in: str, note_in: str, full_in: str) -> Tuple[str, str]:
            blob = glx_normalize_text(" | ".join([str(metric_label_in or ""), str(note_in or ""), str(full_in or "")]))
            parsed = _extract_pbi_guidance_targets_multi(blob, metric_label_in, qd)
            if parsed:
                p0 = parsed[0]
                p_norm = str(p0.get("period_norm") or "").strip()
                p_lbl = str(p0.get("period_label") or "").strip()
                if p_norm or p_lbl:
                    return p_norm, p_lbl or _guidance_period_label_from_norm(p_norm, qd)
            if metric_label_in in {"Revenue guidance", "Adjusted EBIT guidance", "EPS guidance", "FCF target"}:
                fy_year = qd.year + 1 if qd.month == 12 else qd.year
                return f"FY{fy_year}", f"FY {fy_year}"
            if metric_label_in in {"Cost savings target", "Cost savings program", "Cost savings tranche 1", "Cost savings tranche 2"}:
                return "ANNUALIZED_PROGRAM", "Annualized program"
            if metric_label_in == "PB Bank liquidity release":
                return "TIME_ANCHOR", "Time anchor"
            if metric_label_in == "Strategic milestone":
                m_q = re.search(r"\bQ([1-4])\s*(20\d{2})\b|\b(20\d{2})\s*Q([1-4])\b", blob, re.I)
                if m_q:
                    year_txt = m_q.group(2) or m_q.group(3)
                    q_txt = m_q.group(1) or m_q.group(4)
                    if year_txt and q_txt:
                        return f"Q{year_txt}Q{q_txt}", f"Q{q_txt} {year_txt}"
                m_fy = re.search(r"\b(?:fy|fiscal year)\s*(20\d{2})\b|\b(20\d{2})\b", blob, re.I)
                if m_fy:
                    year_txt = m_fy.group(1) or m_fy.group(2)
                    if year_txt:
                        return f"FY{year_txt}", f"FY {year_txt}"
                return "TIME_ANCHOR", "Time anchor"
            return "", ""

        def _note_sentence(metric_label_in: str, target_display_in: str, period_label_in: str, note_in: str) -> str:
            metric_txt = str(metric_label_in or "").strip()
            target_txt = str(target_display_in or "").strip()
            period_txt = str(period_label_in or "").strip()
            note_low = glx_normalize_text(str(note_in or "")).lower()
            prefix = f"{period_txt} {metric_txt}".strip()
            if metric_txt in {"Revenue guidance", "Adjusted EBIT guidance", "EPS guidance", "FCF target"}:
                if "tracking midpoint" in note_low:
                    return f"{prefix} tracking midpoint of {target_txt}.".strip()
                if "tracking low end" in note_low:
                    return f"{prefix} tracking low end of {target_txt}.".strip()
                if "reaffirmed target" in note_low:
                    return f"{prefix} reaffirmed at {target_txt}.".strip()
                if "updated target" in note_low or "raised target" in note_low:
                    return f"{prefix} updated to {target_txt}.".strip()
                return f"{prefix} target {target_txt}.".strip()
            if metric_txt.startswith("Cost savings"):
                if "raised target" in note_low or "updated target" in note_low:
                    return f"{metric_txt} updated to {target_txt}.".strip()
                return f"{metric_txt} target {target_txt}.".strip()
            if metric_txt == "PB Bank liquidity release":
                amt_match = re.search(r"(>=|>|at least)?\s*(\$[0-9]+(?:\.[0-9]+)?[mbn]?)", target_txt, re.I)
                if amt_match:
                    qualifier = str(amt_match.group(1) or "").strip().lower()
                    amount_txt = str(amt_match.group(2) or "").strip()
                    if qualifier in {">=", ">", "at least"}:
                        return f"PB Bank cash optimization target is at least {amount_txt}."
                    return f"PB Bank cash optimization target is {amount_txt}."
                return f"PB Bank cash optimization target is {target_txt}.".strip()
            if metric_txt == "Strategic milestone":
                return glx_normalize_text(str(note_in or "")).strip(" .") + "."
            return glx_normalize_text(str(note_in or "")).strip(" .") + "."

        current_qd: Optional[date] = None
        for rr in range(2, ws_qn.max_row + 1):
            a_val = ws_qn.cell(rr, 1).value
            a_txt = str(a_val or "").strip()
            if a_txt:
                parsed_qd = _qend(a_txt)
                if isinstance(parsed_qd, date):
                    current_qd = parsed_qd
                    continue
            if current_qd != qd:
                continue
            bucket = str(ws_qn.cell(rr, 2).value or "").strip()
            note_txt = glx_normalize_text(str(ws_qn.cell(rr, 3).value or ""))
            metric_label = str(ws_qn.cell(rr, 4).value or "").strip()
            if not note_txt or not metric_label:
                continue
            rows_scanned += 1
            if metric_label not in _pbi_tracker_allowed_labels:
                rejected_bad_metric += 1
                continue
            comment = ws_qn.cell(rr, 3).comment
            full_text = note_txt
            if comment and str(comment.text or "").startswith("Evidence: "):
                raw_comment = str(comment.text or "")
                full_text = glx_normalize_text(raw_comment.split("\n\n", 1)[0].replace("Evidence: ", "", 1))
            target_display = (
                _ui_note_target(note_txt)
                or _extract_pbi_target_display(note_txt, metric_label)
                or _extract_pbi_target_display(full_text, metric_label)
            )
            if metric_label != "Strategic milestone" and not _pbi_target_display_ok(target_display):
                rejected_bad_target += 1
                continue
            parsed_targets = _extract_pbi_guidance_targets_multi(" | ".join([note_txt, full_text]), metric_label, qd)
            period_norm = ""
            period_label = ""
            if parsed_targets:
                for parsed in parsed_targets:
                    if str(parsed.get("label") or "").strip() != metric_label:
                        continue
                    period_norm = str(parsed.get("period_norm") or "").strip()
                    period_label = str(parsed.get("period_label") or "").strip() or _guidance_period_label_from_norm(period_norm, qd)
                    if not target_display:
                        target_display = str(parsed.get("target") or "").strip()
                    break
            if not period_norm and not period_label:
                period_norm, period_label = _default_period(metric_label, note_txt, full_text)
            pair = (
                metric_label.lower(),
                (target_display or note_txt[:120]).lower(),
                str(period_norm or period_label).lower(),
            )
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            tracker_text = _note_sentence(metric_label, target_display, period_label, note_txt)
            guidance_id_map = {
                "Revenue guidance": "rev",
                "Adjusted EBIT guidance": "adj",
                "EPS guidance": "eps",
                "FCF target": "fcf",
                "Cost savings target": "cost_savings",
            }
            guidance_metric_key = guidance_id_map.get(metric_label, "")
            promise_id = (
                f"guidance:{guidance_metric_key}:{period_norm or qd.isoformat()}"
                if guidance_metric_key
                else f"pbi_qn_sheet:{qd.isoformat()}:{metric_label}:{target_display or tracker_text[:48]}"
            )
            rec = {
                "promise_id": promise_id,
                "quarter": qd,
                "metric": metric_label,
                "metric_display": metric_label,
                "period_label": period_label,
                "period_key": period_norm,
                "target_period_norm": period_norm,
                "target_display": target_display,
                "text_full": tracker_text,
                "text_snippet": note_txt,
                "score": 85.0,
                "doc_priority": 8,
                "source": {
                    "source_type": "pbi_quarter_notes_structured",
                    "doc": "Quarter_Notes_UI",
                    "form": "",
                    "section": bucket,
                },
                "guidance_type": "milestone" if metric_label == "Strategic milestone" else "period",
                "has_forward_intent": True,
                "has_time_anchor": True,
                "promise_type": "milestone" if metric_label == "Strategic milestone" else "guidance_range",
                "theme_key": _promise_theme_key(metric_label, full_text or tracker_text, period_norm or period_label),
            }
            rec.update(
                _derive_split_target_meta(
                    metric_label,
                    " | ".join([target_display, full_text or tracker_text]),
                    period_norm or period_label,
                    qd,
                    "quarter_notes_ui",
                    "Quarter_Notes_UI",
                    bucket,
                )
            )
            rec["promise_group"] = str(rec.get("target_group_key") or "")
            out.append(rec)
        if rows_scanned > 0:
            ui_info_rows.append(
                {
                    "quarter": qd,
                    "metric": "Promise_Tracker_UI",
                    "severity": "info",
                    "message": (
                        f"pbi_ui_sheet_rows_scanned={rows_scanned} "
                        f"added={len(out)} "
                        f"rejected_bad_metric={rejected_bad_metric} "
                        f"rejected_bad_target={rejected_bad_target}"
                    ),
                    "source": "Quarter_Notes_UI",
                }
            )
        return out

    def _augment_tracker_with_split_target_totals(
        items: List[Dict[str, Any]],
        qd: date,
    ) -> List[Dict[str, Any]]:
        if not items:
            return items

        def _structured_target_display_for_role(item_local: Dict[str, Any], role_in: str) -> str:
            metric_low = str(item_local.get("metric") or "").strip().lower()
            text_blob = glx_normalize_text(
                " | ".join(
                    [
                        str(item_local.get("target_display") or ""),
                        str(item_local.get("text_full") or item_local.get("text_snippet") or ""),
                    ]
                )
            ).lower()
            period_lbl = str(item_local.get("period_label") or "").strip()
            stage_amt = item_local.get("stage_amount")
            total_amt = item_local.get("program_total_amount")
            increment_amt = item_local.get("increment_amount")
            role_low = str(role_in or "").strip().lower()
            if role_low == "program_total" and total_amt is not None:
                if (
                    metric_low == "cost savings"
                    and increment_amt is not None
                    and (stage_amt is None or float(total_amt) <= max(float(stage_amt), float(increment_amt)) + 1e-6)
                    and re.search(r"\b(additional|second phase|phase 2)\b", text_blob, re.I)
                    and re.search(r"\b(realized|realised|to date|implemented|achieved)\b", text_blob, re.I)
                ):
                    total_amt = max(float(total_amt), float(total_amt) + float(increment_amt))
                if metric_low == "cost savings" or re.search(r"\b(cost savings|cost reduction|annualized savings)\b", text_blob, re.I):
                    return f">= {_fmt_short_money_value_local(float(total_amt))} annualized program"
                if period_lbl:
                    return f">= {_fmt_short_money_value_local(float(total_amt))} in {period_lbl}"
                return f">= {_fmt_short_money_value_local(float(total_amt))} total program"
            if role_low in {"first_tranche", "initial", "phase_1", "additional_tranche", "phase_2", "remaining"} and stage_amt is not None:
                if metric_low == "cost savings" or re.search(r"\b(cost savings|cost reduction|annualized savings)\b", text_blob, re.I):
                    if role_low in {"additional_tranche", "phase_2", "remaining"}:
                        return f">= {_fmt_short_money_value_local(float(stage_amt))} additional annualized savings"
                    return f">= {_fmt_short_money_value_local(float(stage_amt))} annualized savings"
                stage_label_map = {
                    "first_tranche": "initial target",
                    "initial": "initial target",
                    "phase_1": "phase 1 target",
                    "additional_tranche": "additional target",
                    "phase_2": "phase 2 target",
                    "remaining": "remaining target",
                }
                stage_suffix = stage_label_map.get(role_low, "stage target")
                return f">= {_fmt_short_money_value_local(float(stage_amt))} {stage_suffix}"
            return str(item_local.get("target_display") or "").strip()

        expanded_items: List[Dict[str, Any]] = []
        for it in items:
            if not isinstance(it, dict):
                continue
            structure_kind = str(it.get("target_structure_kind") or "").strip().lower()
            stage_kind = str(it.get("stage_kind") or "").strip().lower()
            stage_amt = it.get("stage_amount")
            total_amt = it.get("program_total_amount")
            if structure_kind == "stage_and_total" and stage_kind and stage_amt is not None and total_amt is not None:
                for role_name in (stage_kind, "program_total"):
                    cloned = dict(it)
                    cloned["promise_id"] = hashlib.sha1(
                        f"{it.get('promise_id')}|{role_name}|{qd.isoformat()}".encode("utf-8")
                    ).hexdigest()[:12]
                    cloned["target_structure_kind"] = "stage" if role_name != "program_total" else "program_total"
                    cloned["target_structure_role"] = role_name
                    cloned["stage_kind"] = stage_kind if role_name != "program_total" else ""
                    cloned["target_display"] = _structured_target_display_for_role(cloned, role_name)
                    if str(cloned.get("metric") or "").strip().lower() == "cost savings":
                        if role_name in {"first_tranche", "initial", "phase_1"}:
                            cloned["metric_display"] = "Cost savings tranche 1"
                        elif role_name in {"additional_tranche", "phase_2", "remaining"}:
                            cloned["metric_display"] = "Cost savings tranche 2"
                        else:
                            cloned["metric_display"] = "Cost savings program"
                    else:
                        cloned["metric_display"] = _split_target_metric_display(
                            cloned.get("metric"),
                            " | ".join(
                                [
                                    str(cloned.get("target_display") or ""),
                                    str(cloned.get("text_full") or cloned.get("text_snippet") or ""),
                                ]
                            ),
                            cloned,
                        )
                    expanded_items.append(cloned)
                continue
            expanded_items.append(it)

        out = list(expanded_items)
        existing_keys = {_tracker_identity_key_local(it) for it in out}
        grouped_by_target: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
        for it in out:
            family_key = str(it.get("target_family_key") or "").strip().lower()
            period_key = str(it.get("target_period_key") or it.get("period_key") or "").strip() or "UNK"
            if family_key:
                grouped_by_target.setdefault((family_key, period_key), []).append(it)
        for (family_key, period_key), grp in grouped_by_target.items():
            if any(_split_target_scope_token(it) == "company_total" for it in grp):
                continue
            comp_rows = [it for it in grp if str(it.get("scope_kind") or "").strip().lower().startswith("component")]
            distinct_components = {_split_target_scope_token(it) for it in comp_rows if _split_target_scope_token(it) != "company_total"}
            if len(distinct_components) < 2:
                continue
            if family_key != "advantage_nebraska_45z":
                continue
            target_display = _strong_45z_2026_target_display(
                "45Z-related Adjusted EBITDA in 2026",
                qd,
                "45Z-related Adjusted EBITDA in 2026",
            )
            if not target_display:
                continue
            base = sorted(
                grp,
                key=lambda x: (
                    _source_rank(dict(x.get("source") or {}).get("source_type"), dict(x.get("source") or {}).get("doc")),
                    _text_fragment_penalty(x.get("text_full") or x.get("text_snippet") or ""),
                    -float(x.get("score") or 0.0),
                ),
            )[0]
            scope_labels = [
                str(it.get("metric_display") or it.get("scope_label") or it.get("metric") or "").strip()
                for it in grp
                if str(it.get("metric_display") or it.get("scope_label") or it.get("metric") or "").strip()
            ]
            scope_labels = list(dict.fromkeys(scope_labels))
            summary_bits = ", ".join(scope_labels[:3])
            text_full = f"Company-wide 2026 45Z-related Adjusted EBITDA target. Components include {summary_bits}."
            candidate = {
                "promise_id": hashlib.sha1(f"{qd}|split_total|{family_key}|{period_key}".encode("utf-8")).hexdigest()[:12],
                "metric": "45Z monetization / EBITDA",
                "period_label": str(base.get("period_label") or "FY 2026"),
                "period_key": str(base.get("period_key") or base.get("target_period_norm") or "FY2026"),
                "quarter": qd,
                "text_full": text_full,
                "text_snippet": qn_compact_snippet(text_full, 220),
                "score": max(float(base.get("score") or 0.0), 96.0),
                "source": dict(base.get("source") or {}),
                "source_date": base.get("source_date"),
                "doc_priority": int(base.get("doc_priority") or 0) + 1,
                "reasons": list(dict.fromkeys(list(base.get("reasons") or []) + ["split_target_total_synthesized"])),
                "has_numeric": True,
                "has_time_anchor": True,
                "guidance_type": str(base.get("guidance_type") or "period"),
                "as_of_quarter_end": str(qd),
                "source_doc_end": str(qd),
                "source_filed_date": base.get("source_filed_date"),
                "first_seen_quarter_end": str(base.get("first_seen_quarter_end") or qd),
                "last_seen_quarter_end": str(base.get("last_seen_quarter_end") or qd),
                "referenced_years": [2026],
                "has_forward_intent": True,
                "has_period_anchor": True,
                "target_period_norm": str(base.get("target_period_norm") or base.get("period_key") or "FY2026"),
                "promise_type": str(base.get("promise_type") or "operational"),
                "theme_key": "45z_2026_ebitda",
                "target_display": target_display,
                "_fragment_penalty": 0,
                "_clean_target_bonus": 6,
                "target_family_key": family_key,
                "target_group_key": _split_target_group_key(family_key, period_key, qd),
                "scope_kind": "total",
                "scope_key": "company_total",
                "scope_label": "Company-wide",
                "component_of_total": False,
                "_split_target_amount": 188_000_000.0 if str(ticker or "").upper() == "GPRE" and qd == date(2025, 12, 31) else None,
                "_split_target_rank": 0,
                "promise_group": _split_target_group_key(family_key, period_key, qd),
                "metric_display": "45Z-related Adjusted EBITDA",
            }
            key = _tracker_identity_key_local(candidate)
            if key in existing_keys:
                continue
            out.append(candidate)
            existing_keys.add(key)
        family_rows = [
            it
            for it in out
            if re.search(
                r"\b(45z|advantage nebraska|remaining facilities)\b",
                " ".join(
                    [
                        str(it.get("metric_display") or ""),
                        str(it.get("text_full") or it.get("text_snippet") or ""),
                        str(it.get("target_display") or ""),
                    ]
                ),
                re.I,
            )
        ]
        has_total = any(str(it.get("metric_display") or "").strip() == "45Z-related Adjusted EBITDA" for it in family_rows)
        has_named_component = any(
            "ebitda opportunity" in str(it.get("metric_display") or "").strip().lower()
            or re.search(r"\badvantage nebraska\b", str(it.get("text_full") or it.get("text_snippet") or ""), re.I)
            for it in family_rows
        )
        has_remaining_component = any(
            re.search(r"\bremaining facilities\b", " ".join([str(it.get("metric_display") or ""), str(it.get("text_full") or it.get("text_snippet") or "")]), re.I)
            for it in family_rows
        )
        if not has_total and has_named_component and has_remaining_component:
            base = sorted(
                family_rows,
                key=lambda x: (
                    _source_rank(dict(x.get("source") or {}).get("source_type"), dict(x.get("source") or {}).get("doc")),
                    _text_fragment_penalty(x.get("text_full") or x.get("text_snippet") or ""),
                    -float(x.get("score") or 0.0),
                ),
            )[0]
            target_display = _strong_45z_2026_target_display(
                "45Z-related Adjusted EBITDA in 2026",
                qd,
                "45Z-related Adjusted EBITDA in 2026",
            )
            if target_display:
                group_key = _split_target_group_key("advantage_nebraska_45z", "FY2026", qd)
                candidate = {
                    "promise_id": hashlib.sha1(f"{qd}|split_total_fallback".encode("utf-8")).hexdigest()[:12],
                    "metric": "45Z monetization / EBITDA",
                    "period_label": "FY 2026",
                    "period_key": "FY2026",
                    "quarter": qd,
                    "text_full": "Company-wide 2026 45Z-related Adjusted EBITDA target spanning Nebraska and remaining facilities.",
                    "text_snippet": "Company-wide 2026 45Z-related Adjusted EBITDA target spanning Nebraska and remaining facilities.",
                    "score": 97.0,
                    "source": dict(base.get("source") or {}),
                    "source_date": base.get("source_date"),
                    "doc_priority": int(base.get("doc_priority") or 0) + 1,
                    "reasons": list(dict.fromkeys(list(base.get("reasons") or []) + ["split_target_total_fallback"])),
                    "has_numeric": True,
                    "has_time_anchor": True,
                    "guidance_type": "period",
                    "as_of_quarter_end": str(qd),
                    "source_doc_end": str(qd),
                    "source_filed_date": base.get("source_filed_date"),
                    "first_seen_quarter_end": str(base.get("first_seen_quarter_end") or qd),
                    "last_seen_quarter_end": str(base.get("last_seen_quarter_end") or qd),
                    "referenced_years": [2026],
                    "has_forward_intent": True,
                    "has_period_anchor": True,
                    "target_period_norm": "FY2026",
                    "promise_type": "operational",
                    "theme_key": "45z_2026_ebitda",
                    "target_display": target_display,
                    "_fragment_penalty": 0,
                    "_clean_target_bonus": 6,
                    "target_family_key": "advantage_nebraska_45z",
                    "target_group_key": group_key,
                    "scope_kind": "total",
                    "scope_key": "company_total",
                    "scope_label": "Company-wide",
                    "component_of_total": False,
                    "_split_target_amount": 188_000_000.0 if str(ticker or "").upper() == "GPRE" and qd == date(2025, 12, 31) else None,
                    "_split_target_rank": 0,
                    "promise_group": group_key,
                    "metric_display": "45Z-related Adjusted EBITDA",
                }
                key = _tracker_identity_key_local(candidate)
                if key not in existing_keys:
                    out.append(candidate)
                    existing_keys.add(key)
        return out

    slide_signals = _load_profile_slide_signals()
    if slide_signals:
        for qd in q_window:
            existing_by_metric: Dict[Tuple[str, str, str], int] = {}
            for idx_item, it in enumerate(grouped.get(qd, [])):
                m_existing = _tracker_identity_key_local(it)
                if m_existing and m_existing not in existing_by_metric:
                    existing_by_metric[m_existing] = idx_item
            existing_texts = {
                glx_normalize_text(str(it.get("text_full") or it.get("text_snippet") or "")).lower()
                for it in grouped.get(qd, [])
                if str(it.get("text_full") or it.get("text_snippet") or "").strip()
            }
            for rec in sorted(
                [x for x in slide_signals if x.get("quarter") == qd],
                key=lambda z: -float(z.get("score") or 0.0),
            ):
                metric_name = str(rec.get("metric") or "").strip()
                txt_full = glx_normalize_text(str(rec.get("text") or ""))
                if not metric_name or not txt_full:
                    continue
                if metric_name.lower() in {"management target", "utilization", "risk management"} and not bool(rec.get("is_numeric_target")) and not bool(rec.get("is_milestone")):
                    continue
                if txt_full.lower() in existing_texts or _slide_signal_noise(txt_full):
                    continue
                period_label = ""
                m_year = re.search(r"\b(20\d{2})\b", txt_full)
                if m_year:
                    period_label = f"FY {m_year.group(1)}"
                candidate_row = {
                    "promise_id": hashlib.sha1(f"{qd}|tracker|{metric_name}|{txt_full}".encode("utf-8")).hexdigest()[:12],
                    "quarter": qd,
                    "metric": metric_name,
                    "text_snippet": qn_compact_snippet(txt_full, 220),
                    "text_full": txt_full,
                    "period_label": period_label,
                    "guidance_type": "period" if bool(rec.get("is_numeric_target")) else ("milestone" if bool(rec.get("is_milestone")) else ""),
                    "first_seen_quarter_end": str(qd),
                    "last_seen_quarter_end": str(qd),
                    "source": {
                        "source_type": str(rec.get("source_type") or "earnings_presentation"),
                        "doc": str(rec.get("source_doc") or ""),
                        "form": "presentation",
                    },
                    "score": float(rec.get("score") or 0.0),
                    "doc_priority": 8,
                    "reasons": ["slides_signal"],
                    "theme_key": str(rec.get("theme_key") or _promise_theme_key(metric_name, txt_full, period_label)),
                    "target_display": str(rec.get("target_display") or ""),
                    "_fragment_penalty": _text_fragment_penalty(txt_full),
                    "_clean_target_bonus": _clean_target_bonus(txt_full),
                }
                candidate_row.update(
                    _derive_split_target_meta(
                        metric_name,
                        txt_full,
                        period_label,
                        qd,
                        rec.get("source_type") or "earnings_presentation",
                        rec.get("source_doc") or "",
                        "",
                    )
                )
                candidate_row["promise_group"] = str(candidate_row.get("target_group_key") or "")
                candidate_row["metric_display"] = str(rec.get("metric_display") or _split_target_metric_display(metric_name, txt_full, candidate_row))
                metric_key = _tracker_identity_key_local(candidate_row)
                if metric_key in existing_by_metric:
                    existing_item = grouped.setdefault(qd, [])[existing_by_metric[metric_key]]
                    prefer_candidate = _promise_quality_key(candidate_row) < _promise_quality_key(existing_item)
                    if (
                        str(candidate_row.get("target_family_key") or "").strip().lower() == "advantage_nebraska_45z"
                        and str(candidate_row.get("scope_kind") or "").strip().lower().startswith("component")
                        and "ebitda opportunity" in txt_full.lower()
                    ):
                        prefer_candidate = True
                    if prefer_candidate:
                        grouped[qd][existing_by_metric[metric_key]] = candidate_row
                        existing_texts.add(txt_full.lower())
                    continue
                grouped.setdefault(qd, []).append(candidate_row)
                existing_by_metric[metric_key] = len(grouped.get(qd, [])) - 1
                existing_texts.add(txt_full.lower())
            grouped[qd] = sorted(
                [x for x in grouped.get(qd, []) if _promise_tracker_keep_item(x)],
                key=lambda x: (
                    metric_priority.get(str(x.get("metric") or FORWARD_NOTES_LABEL), 99),
                    _promise_quality_key(x),
                    -float(x.get("score") or 0.0),
                    -int(x.get("doc_priority") or 0),
                ),
            )[:15]

    tracker_priority_metrics = {"45Z monetization / EBITDA", "Cost savings", "Debt reduction", "Strategic milestone"}
    for qd in q_window:
        current_items = list(grouped.get(qd, []))
        existing_by_metric = {
            _tracker_identity_key_local(it): idx
            for idx, it in enumerate(current_items)
            if str(it.get("metric") or "").strip() or str(it.get("scope_key") or "").strip()
        }
        additions = 0
        priority_additions = 0
        for note_item in quarter_note_rows_map.get(qd, []) or []:
            candidate_row = _tracker_candidate_from_qnote(qd, note_item)
            if candidate_row is None or not _promise_tracker_keep_item(candidate_row):
                continue
            metric_name = str(candidate_row.get("metric") or "").strip()
            metric_key = _tracker_identity_key_local(candidate_row)
            if not any(metric_key):
                continue
            is_priority_metric = metric_name in tracker_priority_metrics
            if metric_key in existing_by_metric:
                existing_item = current_items[existing_by_metric[metric_key]]
                if _promise_quality_key(candidate_row) < _promise_quality_key(existing_item):
                    current_items[existing_by_metric[metric_key]] = candidate_row
                continue
            if not is_priority_metric and len(current_items) >= 3:
                continue
            current_items.append(candidate_row)
            existing_by_metric[metric_key] = len(current_items) - 1
            additions += 1
            if is_priority_metric:
                priority_additions += 1
            if len(current_items) >= 5 and priority_additions >= 2:
                break
        if not is_pbi_profile:
            existing_metric_names = {
                str(it.get("metric") or "").strip().lower()
                for it in current_items
                if str(it.get("metric") or "").strip()
            }
            targeted_additions = 0
            for wanted_metric in ["45Z monetization / EBITDA", "Debt reduction", "Cost savings", "Strategic milestone"]:
                if wanted_metric.lower() in existing_metric_names:
                    continue
                best_candidate = None
                best_rank = None
                for note_item in quarter_note_rows_map.get(qd, []) or []:
                    candidate_row = _tracker_candidate_from_qnote(qd, note_item)
                    if candidate_row is None:
                        continue
                    if str(candidate_row.get("metric") or "").strip() != wanted_metric:
                        continue
                    rank = (
                        _promise_quality_key(candidate_row),
                        -float(candidate_row.get("score") or 0.0),
                        -int(candidate_row.get("doc_priority") or 0),
                    )
                    if best_rank is None or rank < best_rank:
                        best_rank = rank
                        best_candidate = candidate_row
                if best_candidate is None:
                    continue
                metric_key = _tracker_identity_key_local(best_candidate)
                if any(metric_key) and metric_key not in existing_by_metric:
                    current_items.append(best_candidate)
                    existing_by_metric[metric_key] = len(current_items) - 1
                    existing_metric_names.add(wanted_metric.lower())
                    targeted_additions += 1
            additions += targeted_additions
        if is_gpre_profile and not current_items:
            gpre_rescue_added = 0
            for note_item in quarter_note_rows_map.get(qd, []) or []:
                txt_full = glx_normalize_text(str(note_item.get("text_full") or note_item.get("comment_full_text") or ""))
                if not txt_full:
                    continue
                metric_display = str(note_item.get("_metric_display") or note_item.get("metric_canon") or note_item.get("metric_tag") or "").strip()
                metric_name = ""
                blob = " | ".join([metric_display, txt_full])
                if re.search(r"\b45z\b|tax credit monetization|qualify for production tax credits", blob, re.I):
                    metric_name = "45Z monetization / EBITDA"
                elif re.search(r"\b(debt reduction|repay|repaid|repayment|sale of obion|obion)\b", blob, re.I):
                    metric_name = "Debt reduction"
                elif re.search(r"\b(cost savings|cost reduction|reorganization|expense reduction)\b", blob, re.I):
                    metric_name = "Cost savings"
                elif re.search(r"\b(fully operational|online and ramping|agreement executed|construction|commissioning|milestone)\b", blob, re.I):
                    metric_name = "Strategic milestone"
                if not metric_name:
                    continue
                source = dict(note_item.get("source") or {})
                source_type = str(source.get("source_type") or source.get("doc_type") or "").lower()
                if not (preferred_promise_source_re.search(source_type) or any(k in source_type for k in ("presentation", "release", "transcript", "ceo"))):
                    continue
                candidate_row = {
                    "promise_id": str(note_item.get("note_id") or hashlib.sha1(f"{qd}|gpre_visible_qnote_recall|{txt_full}".encode("utf-8")).hexdigest()[:12]),
                    "quarter": qd,
                    "metric": metric_name,
                    "text_snippet": qn_compact_snippet(txt_full, 220),
                    "text_full": txt_full,
                    "period_label": f"FY {re.search(r'(20\d{2})', txt_full).group(1)}" if re.search(r'(20\d{2})', txt_full) else "",
                    "guidance_type": "milestone" if metric_name == "Strategic milestone" else "period",
                    "first_seen_quarter_end": str(qd),
                    "last_seen_quarter_end": str(qd),
                    "source": {
                        "source_type": str(source.get("source_type") or source.get("doc_type") or "quarter_notes_ui"),
                        "doc": str(source.get("doc") or "Quarter_Notes_UI"),
                        "form": str(source.get("form") or ""),
                        "section": str(source.get("section") or ""),
                    },
                    "score": float(note_item.get("score") or 0.0) + 3.0,
                    "doc_priority": 7,
                    "reasons": ["quarter_notes_ui_visible_recall"],
                    "theme_key": str(note_item.get("theme_key") or _promise_theme_key(metric_name, txt_full, "")),
                    "_fragment_penalty": _text_fragment_penalty(txt_full),
                    "_clean_target_bonus": _clean_target_bonus(txt_full),
                    "target_display": _tracker_target_display_from_qnote(qd, metric_name, txt_full),
                }
                if metric_name == "Debt reduction":
                    candidate_row["status_hint"] = "completed"
                    candidate_row["latest_display"] = "Debt repaid"
                elif metric_name == "Strategic milestone":
                    if re.search(r"\bfully operational\b", txt_full, re.I):
                        candidate_row["status_hint"] = "completed"
                        candidate_row["latest_display"] = "Fully operational"
                    elif re.search(r"\bonline and ramping\b", txt_full, re.I):
                        candidate_row["status_hint"] = "in progress"
                        candidate_row["latest_display"] = "Online and ramping"
                candidate_row.update(
                    _derive_split_target_meta(
                        metric_name,
                        txt_full,
                        candidate_row.get("guidance_type") or "",
                        qd,
                        source.get("source_type") or source.get("doc_type") or "quarter_notes_ui",
                        source.get("doc") or "Quarter_Notes_UI",
                        source.get("section") or "",
                    )
                )
                candidate_row["promise_group"] = str(candidate_row.get("target_group_key") or "")
                candidate_row["metric_display"] = _split_target_metric_display(metric_name, txt_full, candidate_row)
                current_items.append(candidate_row)
                gpre_rescue_added += 1
                if gpre_rescue_added >= 4:
                    break
            additions += gpre_rescue_added
        if additions > 0:
            grouped[qd] = sorted(
                current_items,
                key=lambda x: (
                    metric_priority.get(str(x.get("metric") or FORWARD_NOTES_LABEL), 99),
                    _promise_quality_key(x),
                    -float(x.get("score") or 0.0),
                    -int(x.get("doc_priority") or 0),
                ),
            )[:15]
            ui_info_rows.append(
                {
                    "quarter": qd,
                    "metric": "Promise_Tracker_UI",
                    "severity": "info",
                    "message": f"quarter_notes_ui_recall added={additions}",
                    "source": "Quarter_Notes_UI",
                }
            )

    for qd in list(grouped.keys()):
        grouped[qd] = sorted(
            _augment_tracker_with_split_target_totals(grouped.get(qd, []), qd),
            key=lambda x: (
                metric_priority.get(str(x.get("metric") or FORWARD_NOTES_LABEL), 99),
                int(x.get("_split_target_rank") or 9),
                _promise_quality_key(x),
                -float(x.get("score") or 0.0),
                -int(x.get("doc_priority") or 0),
            ),
        )[:15]
        deduped_tracker_rows: List[Dict[str, Any]] = []
        seen_tracker_keys: set[Tuple[str, str]] = set()
        for item in grouped[qd]:
            dedup_key = (
                str(item.get("metric_display") or item.get("metric") or "").strip().lower(),
                glx_normalize_text(str(item.get("target_display") or item.get("text_full") or item.get("text_snippet") or "")).lower(),
            )
            if dedup_key in seen_tracker_keys:
                continue
            seen_tracker_keys.add(dedup_key)
            deduped_tracker_rows.append(item)
        grouped[qd] = deduped_tracker_rows[:15]
        if render_visible and is_pbi_profile:
            final_items: List[Dict[str, Any]] = []
            seen_keys: set[Tuple[str, str]] = set()
            for it in grouped[qd]:
                if not _pbi_final_tracker_keep_item(it):
                    ui_info_rows.append(
                        {
                            "quarter": qd,
                            "metric": "Promise_Tracker_UI",
                            "severity": "info",
                            "message": "dropped_reason=pbi_final_tracker_filter",
                            "source": str(dict(it.get("source") or {}).get("doc") or ""),
                        }
                    )
                    continue
                dedup_key = (
                    str(it.get("metric_display") or it.get("metric") or "").strip().lower(),
                    str(it.get("target_display") or "").strip().lower(),
                )
                if dedup_key in seen_keys:
                    continue
                seen_keys.add(dedup_key)
                final_items.append(it)
            grouped[qd] = final_items[:8]
        elif render_visible and is_gpre_profile:
            grouped[qd] = [it for it in grouped[qd] if _gpre_final_tracker_keep_item(it)]

    if is_pbi_profile and isinstance(gstore, dict):
        for qd in q_window:
            existing_items = list(grouped.get(qd, []))
            existing_metric_keys = {
                str(it.get("metric_display") or it.get("metric") or "").strip().lower()
                for it in existing_items
                if str(it.get("metric_display") or it.get("metric") or "").strip()
            }
            guidance_items = gstore.get(str(qd), []) or []
            added = 0
            for it in guidance_items:
                if not isinstance(it, dict):
                    continue
                metric_name = str(it.get("metric") or "").strip()
                if metric_name in {"", FORWARD_NOTES_LABEL, "Other", "Unknown"}:
                    continue
                metric_label = _classify_pbi_metric_label(metric_name, metric_name) or metric_name
                if metric_label not in _pbi_tracker_allowed_labels:
                    continue
                target_display = _guidance_value_snip(it)
                if not _pbi_target_display_ok(target_display):
                    continue
                dedup_metric = metric_label.strip().lower()
                if dedup_metric in existing_metric_keys:
                    continue
                period_norm = str(it.get("target_period_norm") or it.get("period_norm") or "UNK").strip() or "UNK"
                period_label = _guidance_period_label_from_norm(period_norm, qd)
                src_it = dict(it.get("source") or {})
                tracker_row = {
                    "promise_id": f"guidance_tracker:{qd.isoformat()}:{metric_label}:{period_norm}",
                    "quarter": qd,
                    "metric": metric_label,
                    "metric_display": metric_label,
                    "period_label": period_label,
                    "period_key": period_norm,
                    "target_period_norm": period_norm,
                    "target_display": target_display,
                    "text_full": f"{period_label} {metric_label} {target_display}".strip(),
                    "text_snippet": f"{period_label} {metric_label} {target_display}".strip(),
                    "score": float(it.get("score") or 82.0),
                    "doc_priority": int(it.get("source_priority") or 0),
                    "source": {
                        "source_type": str(src_it.get("source_type") or "guidance_snapshot"),
                        "doc": str(src_it.get("doc") or ""),
                        "form": str(src_it.get("form") or ""),
                        "section": str(src_it.get("section") or src_it.get("section_or_page") or ""),
                    },
                    "guidance_type": str(it.get("guidance_type") or "period"),
                    "has_forward_intent": True,
                    "has_time_anchor": True,
                    "promise_type": "guidance_range",
                    "theme_key": _promise_theme_key(metric_label, target_display, period_norm),
                }
                existing_items.append(tracker_row)
                existing_metric_keys.add(dedup_metric)
                added += 1
            if added > 0:
                grouped[qd] = existing_items[:8]
                ui_info_rows.append(
                    {
                        "quarter": qd,
                        "metric": "Promise_Tracker_UI",
                        "severity": "info",
                        "message": f"guidance_snapshot_fallback_added count={added}",
                        "source": "guidance_snapshot",
                    }
                )

    if is_pbi_profile and isinstance(slides_guidance, pd.DataFrame) and not slides_guidance.empty:
        sg = slides_guidance.copy()
        if "quarter" in sg.columns:
            sg["quarter_norm"] = pd.to_datetime(sg["quarter"], errors="coerce").dt.date
        else:
            sg["quarter_norm"] = pd.NaT
        for qd in q_window:
            existing_items = list(grouped.get(qd, []))
            existing_metric_keys = {
                str(it.get("metric_display") or it.get("metric") or "").strip().lower()
                for it in existing_items
                if str(it.get("metric_display") or it.get("metric") or "").strip()
            }
            try:
                sg_rows = sg.loc[sg["quarter_norm"] == qd].to_dict("records")
            except Exception:
                sg_rows = []
            added = 0
            seen_label_target: set[Tuple[str, str]] = set()
            for rec in sg_rows:
                txt_full = glx_normalize_text(str(rec.get("line") or ""))
                if not txt_full or _slide_signal_noise(txt_full) or _looks_pbi_fragment_text(txt_full):
                    continue
                metric_hint = str(rec.get("metric_hint") or "").strip()
                metric_label = _classify_pbi_metric_label(f"{metric_hint} | {txt_full}", metric_hint)
                if metric_label not in _pbi_tracker_allowed_labels:
                    continue
                target_display = _extract_pbi_target_display(txt_full, metric_label)
                if not _pbi_target_display_ok(target_display):
                    continue
                dedup_metric = metric_label.strip().lower()
                dedup_pair = (dedup_metric, target_display.strip().lower())
                if dedup_metric in existing_metric_keys or dedup_pair in seen_label_target:
                    continue
                year_guess = qd.year + 1 if re.search(r"\b(fy\s*20\d{2}|full[- ]?year)\b", txt_full, re.I) and qd.month == 12 else qd.year
                m_fy = re.search(r"\bfy\s*(20\d{2})\b", txt_full, re.I)
                if m_fy:
                    year_guess = int(m_fy.group(1))
                period_norm = f"FY{year_guess}" if re.search(r"\b(fy\s*20\d{2}|full[- ]?year)\b", txt_full, re.I) else "TIME_ANCHOR"
                period_label = _guidance_period_label_from_norm(period_norm, qd) or f"Q{((qd.month - 1) // 3) + 1} {qd.year}"
                tracker_row = {
                    "promise_id": f"slides_guidance_tracker:{qd.isoformat()}:{metric_label}:{target_display}",
                    "quarter": qd,
                    "metric": metric_label,
                    "metric_display": metric_label,
                    "period_label": period_label,
                    "period_key": period_norm,
                    "target_period_norm": period_norm,
                    "target_display": target_display,
                    "text_full": txt_full,
                    "text_snippet": qn_compact_snippet(txt_full, 220),
                    "score": 80.0,
                    "doc_priority": 7,
                    "source": {
                        "source_type": "earnings_presentation",
                        "doc": str(rec.get("doc") or ""),
                        "form": "presentation",
                        "section": f"page {rec.get('page')}" if rec.get("page") is not None else "",
                    },
                    "guidance_type": "period",
                    "has_forward_intent": True,
                    "has_time_anchor": True,
                    "promise_type": "guidance_range",
                    "theme_key": _promise_theme_key(metric_label, txt_full, period_norm),
                }
                existing_items.append(tracker_row)
                existing_metric_keys.add(dedup_metric)
                seen_label_target.add(dedup_pair)
                added += 1
            if added > 0:
                grouped[qd] = existing_items[:8]
                ui_info_rows.append(
                    {
                        "quarter": qd,
                        "metric": "Promise_Tracker_UI",
                        "severity": "info",
                        "message": f"slides_guidance_fallback_added count={added}",
                        "source": "Slides_Guidance",
                    }
                )

    if render_visible and is_pbi_profile:
        def _pbi_tracker_metric_label(item_in: Dict[str, Any]) -> str:
            return _classify_pbi_metric_label(
                " | ".join(
                    [
                        str(item_in.get("metric_display") or ""),
                        str(item_in.get("metric") or ""),
                        str(item_in.get("text_full") or item_in.get("text_snippet") or ""),
                        str(item_in.get("target_display") or ""),
                    ]
                ),
                str(item_in.get("metric_display") or item_in.get("metric") or ""),
            )

        def _pbi_tracker_target_key(item_in: Dict[str, Any]) -> str:
            txt_local = str(item_in.get("target_display") or "").strip().lower()
            txt_local = txt_local.replace(",", "")
            txt_local = re.sub(r"\s+", " ", txt_local)
            txt_local = re.sub(r"\bannualized program\b", "annualized", txt_local)
            metric_label = str(item_in.get("metric_display") or item_in.get("metric") or "").strip().lower()
            if metric_label.startswith("cost savings"):
                money_targets = _extract_money_targets_for_display(txt_local)
                if money_targets:
                    txt_local = "|".join(f"{float(v):.1f}" for v in money_targets[:2])
                else:
                    txt_local = re.sub(r"\bannualized savings\b", "", txt_local)
                    txt_local = re.sub(r"\bannualized\b", "", txt_local)
                    txt_local = re.sub(r"\bsavings\b", "", txt_local)
                    txt_local = re.sub(r"\s+", " ", txt_local).strip()
            return txt_local

        for qd in q_window:
            ui_rows = _build_pbi_tracker_rows_from_ui_sheet(qd)
            authoritative_items: List[Dict[str, Any]] = [dict(x) for x in ui_rows]
            for rec in list(grouped.get(qd, [])):
                metric_label = _pbi_tracker_metric_label(rec)
                if metric_label != "Strategic milestone":
                    continue
                if _pbi_final_tracker_keep_item(rec):
                    authoritative_items.append(dict(rec))
            if not authoritative_items:
                existing_items = []
                for rec in _build_pbi_tracker_ui_rows(qd):
                    existing_items.append(rec)
                for rec in _build_pbi_tracker_fallback_rows(qd):
                    existing_items.append(rec)
                for rec in _build_pbi_qnote_tracker_fallback_rows(qd):
                    existing_items.append(rec)
                authoritative_items = [x for x in existing_items if _pbi_final_tracker_keep_item(x)]

            best_by_pair: Dict[Tuple[str, str], Dict[str, Any]] = {}
            best_rank: Dict[Tuple[str, str], Tuple[Any, ...]] = {}
            for item in authoritative_items:
                metric_label = _pbi_tracker_metric_label(item)
                if metric_label not in _pbi_tracker_allowed_labels:
                    continue
                item["metric_display"] = metric_label
                if not _pbi_final_tracker_keep_item(item):
                    continue
                src_type = str(dict(item.get("source") or {}).get("source_type") or "").strip().lower()
                pair = (metric_label.lower(), _pbi_tracker_target_key(item))
                has_period = 0 if str(item.get("period_label") or item.get("target_period_norm") or "").strip() else 1
                rank = (
                    0 if src_type == "pbi_quarter_notes_structured" else 1,
                    has_period,
                    _promise_quality_key(item),
                    -float(item.get("score") or 0.0),
                    -int(item.get("doc_priority") or 0),
                )
                if pair not in best_rank or rank < best_rank[pair]:
                    best_rank[pair] = rank
                    best_by_pair[pair] = item
            grouped[qd] = sorted(
                list(best_by_pair.values()),
                key=lambda x: (
                    metric_priority.get(str(x.get("metric_display") or x.get("metric") or FORWARD_NOTES_LABEL), 99),
                    -max(_extract_money_targets_for_display(str(x.get("target_display") or "")) or [0.0]),
                    int(x.get("_split_target_rank") or 9),
                    _promise_quality_key(x),
                    -float(x.get("score") or 0.0),
                    -int(x.get("doc_priority") or 0),
                ),
            )[:8]
            metric_caps = {
                "Revenue guidance": 1,
                "Adjusted EBIT guidance": 1,
                "EPS guidance": 1,
                "FCF target": 1,
                "PB Bank liquidity release": 1,
                "Deleveraging target": 1,
                "SendTech / Presort operating target": 1,
                "Strategic milestone": 1,
            }
            capped_rows: List[Dict[str, Any]] = []
            metric_counts: Dict[str, int] = {}
            for item in grouped[qd]:
                label = str(item.get("metric_display") or item.get("metric") or "").strip()
                cap = metric_caps.get(label, 2 if label.startswith("Cost savings") else 1)
                used = metric_counts.get(label, 0)
                if used >= cap:
                    continue
                metric_counts[label] = used + 1
                capped_rows.append(item)
            grouped[qd] = capped_rows[:8]
            if ui_rows:
                ui_info_rows.append(
                    {
                        "quarter": qd,
                        "metric": "Promise_Tracker_UI",
                        "severity": "info",
                        "message": f"pbi_ui_sheet_tracker_selected count={len(grouped[qd])}",
                        "source": "Quarter_Notes_UI",
                    }
                )

    _store_tracker_state(grouped, q_window)
    if not render_visible or ws is None:
        return qa_rows

    ws["A1"] = f"Generated at {ts} | Quarter list view | high-signal promise mode"
    ws["A1"].font = Font(bold=True, size=header_size)

    ev_map_q: Dict[Tuple[str, date], int] = {}
    ev_map_pid: Dict[str, int] = {}
    if promise_evidence_df is not None and not promise_evidence_df.empty:
        for i, rr in promise_evidence_df.iterrows():
            pid = str(rr.get("promise_id") or "").strip()
            qd = _qend(rr.get("quarter"))
            if not pid:
                continue
            if qd is not None:
                ev_map_q[(pid, qd)] = i + 2
            if pid not in ev_map_pid:
                ev_map_pid[pid] = i + 2

    hdr_fill = PatternFill("solid", fgColor="F2F2F2")
    sep_side = Side(style="medium")
    row_idx = 2
    promise_rows_for_progress: List[Dict[str, Any]] = []

    for qd in q_window:
        h = ws.cell(row=row_idx, column=1, value=str(qd))
        h.font = Font(bold=True, size=header_size)
        h.fill = hdr_fill
        h.alignment = Alignment(horizontal="left", vertical="center")
        for cc in (2, 3, 4):
            ws.cell(row=row_idx, column=cc, value=None).fill = hdr_fill
        row_idx += 1

        items = grouped.get(qd, [])
        if not items:
            row_idx += 1
            continue

        for rec in items:
            pid = str(rec.get("promise_id") or "")
            metric_label = _display_tracker_metric(rec)
            metric_prefix = f"[{metric_label}] " if metric_label and metric_label != FORWARD_NOTES_LABEL else ""
            target_disp = str(rec.get("target_display") or "").strip()
            if is_pbi_profile and not target_disp:
                target_disp = _extract_pbi_target_display(
                    " | ".join(
                        [
                            str(rec.get("text_full") or rec.get("text_snippet") or ""),
                            str(rec.get("metric_display") or ""),
                            str(rec.get("metric") or ""),
                        ]
                    ),
                    metric_label,
                )
            theme_key = str(rec.get("theme_key") or "").strip().lower()
            structure_kind = str(rec.get("target_structure_kind") or "").strip().lower()
            snippet_txt = str(rec.get("text_snippet") or "").strip()
            if (
                is_pbi_profile
                and metric_label == "Strategic milestone"
                and target_disp.lower() in {"", "milestone"}
            ):
                txt = f"{metric_prefix}{snippet_txt}".strip()
            elif target_disp and (
                is_pbi_profile
                or theme_key.startswith("45z_2026_ebitda")
                or structure_kind in {"stage", "program_total", "stage_and_total"}
            ):
                txt = f"{metric_prefix}{target_disp}".strip()
                if (
                    snippet_txt
                    and not is_pbi_profile
                    and glx_normalize_text(target_disp).lower() in glx_normalize_text(snippet_txt).lower()
                ):
                    txt = f"{txt} | {snippet_txt}".strip()
            else:
                txt = f"{metric_prefix}{snippet_txt}".strip()
            period_label = str(rec.get("period_label") or "").strip()
            if period_label and period_label.lower() not in {"unknown", "n/a"}:
                txt = f"{txt} | {period_label}" if txt else period_label
            gtype = str(rec.get("guidance_type") or "").strip().lower()
            fs_q = str(rec.get("first_seen_quarter_end") or "").strip()
            ls_q = str(rec.get("last_seen_quarter_end") or "").strip()
            if gtype in {"run-rate", "ongoing"} and fs_q:
                fs_lbl = _quarter_lbl(fs_q)
                ls_lbl = _quarter_lbl(ls_q) if ls_q else fs_lbl
                stated = f"stated {fs_lbl} | last seen {ls_lbl}"
                txt = f"{txt} | {stated}" if txt else stated

            c_txt = ws.cell(row=row_idx, column=2, value=txt)
            c_txt.alignment = Alignment(wrap_text=True, vertical="top")
            c_txt.font = Font(size=13, color="000000")

            c_metric = ws.cell(row=row_idx, column=3, value=metric_label if metric_label != FORWARD_NOTES_LABEL else "")
            c_metric.alignment = Alignment(vertical="top")
            c_metric.font = Font(size=11, color="000000", bold=True)

            src = dict(rec.get("source") or {})
            src_bits = [
                f"Source: {src.get('source_type') or 'n/a'}",
                f"form={src.get('form') or ''}",
                f"accn={src.get('accn') or ''}",
                f"doc={src.get('doc') or ''}",
                f"section={src.get('section') or ''}",
                f"score={float(rec.get('score') or 0.0):.1f}",
                f"reasons={','.join(list(rec.get('reasons') or [])[:6])}",
                f"guidance_type={rec.get('guidance_type') or ''}",
                f"first_seen={rec.get('first_seen_quarter_end') or ''}",
                f"last_seen={rec.get('last_seen_quarter_end') or ''}",
            ]
            comment_txt = f"Evidence: {rec.get('text_full') or ''}\\n\\n" + " | ".join([x for x in src_bits if x])
            try:
                _set_cell_comment_local(c_txt, comment_txt)
            except Exception:
                pass

            if pid:
                promise_rows_for_progress.append({"promise_id": pid})
                er = ev_map_q.get((pid, qd)) or ev_map_pid.get(pid)
                if er is not None:
                    link_cell = ws.cell(row=row_idx, column=4, value="source")
                    _apply_hyperlink_look(link_cell, f"#'Promise_Evidence'!A{er}")
                    link_cell.alignment = Alignment(horizontal="left", vertical="top")
                    link_cell.font = Font(size=11, color="0563C1", underline="single")

            segs = max(1, (len(txt) + 179) // 180)
            ws.row_dimensions[row_idx].height = max(18, min(56, 18 + 8 * segs))
            row_idx += 1

        sep_row = row_idx - 1
        for cc in (1, 2, 3, 4):
            cell = ws.cell(row=sep_row, column=cc)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=sep_side,
            )

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 233.57
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    if promise_rows_for_progress:
        ui_state["promise_rows"] = pd.DataFrame(promise_rows_for_progress).drop_duplicates(["promise_id"]).reset_index(drop=True)
    if milestone_suppressed_count > 0:
        ui_info_rows.append(
            {
                "quarter": q_window[0] if q_window else None,
                "metric": "Promise_Tracker_UI",
                "severity": "info",
                "message": f"milestone_suppressed_ui count={int(milestone_suppressed_count)}",
                "source": "pipeline",
            }
        )
    return qa_rows
