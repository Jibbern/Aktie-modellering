"""Quarter narrative record building and workbook surface rendering."""
from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from datetime import date, timedelta
from typing import Any, Dict, List, Mapping, Optional, Sequence, Set, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class _QuarterNarrativeFiscalPeriodProfile:
    year_end_month: int = 12
    year_end_day: int = 31
    year_label: str = "end"


def _quarter_narrative_safe_date(year: int, month: int, day: int) -> date:
    month = max(1, min(12, int(month)))
    day = max(1, min(31, int(day)))
    while True:
        try:
            return date(int(year), month, day)
        except ValueError:
            day -= 1


def _quarter_narrative_fiscal_profile_from_workbook(
    wb: Optional[Workbook],
    ticker: Any = "",
    fiscal_profile: Any = None,
) -> _QuarterNarrativeFiscalPeriodProfile:
    """Resolve fiscal-year-end behavior for quarter narrative labels."""

    def _profile(month: Any, day: Any, label: Any = "") -> _QuarterNarrativeFiscalPeriodProfile:
        m = int(month or 12)
        d = int(day or 31)
        mode = str(label or "").strip().lower()
        if mode not in {"start", "end"}:
            mode = "start" if m <= 2 else "end"
        return _QuarterNarrativeFiscalPeriodProfile(m, d, mode)

    if isinstance(fiscal_profile, _QuarterNarrativeFiscalPeriodProfile):
        return fiscal_profile
    if isinstance(fiscal_profile, Mapping):
        month = fiscal_profile.get("year_end_month") or fiscal_profile.get("fiscal_year_end_month")
        day = fiscal_profile.get("year_end_day") or fiscal_profile.get("fiscal_year_end_day")
        if month and day:
            return _profile(month, day, fiscal_profile.get("year_label") or fiscal_profile.get("fiscal_year_label"))
    if isinstance(fiscal_profile, (tuple, list)) and len(fiscal_profile) >= 2:
        return _profile(fiscal_profile[0], fiscal_profile[1], fiscal_profile[2] if len(fiscal_profile) > 2 else "")

    if wb is not None:
        for sheet_name in ("SUMMARY", "Summary", "Model_Info", "QA_Checks"):
            if sheet_name not in getattr(wb, "sheetnames", []):
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=min(int(ws.max_row or 0), 80), min_col=1, max_col=min(int(ws.max_column or 0), 10), values_only=True):
                blob = " ".join(str(v) for v in row if v not in (None, ""))
                if not blob:
                    continue
                m = re.search(r"\b(?:FY|fiscal year|year)\s*end(?:ed)?\s*(?:\(|:)?\s*(20\d{2})-(\d{1,2})-(\d{1,2})", blob, re.I)
                if m:
                    return _profile(m.group(2), m.group(3), "")
                m = re.search(r"\b(?:FY|fiscal year|year)\s*end(?:ed)?\s*(?:\(|:)?\s*([A-Za-z]+)\s+(\d{1,2})", blob, re.I)
                if m:
                    try:
                        month = pd.to_datetime(m.group(1), format="%B", errors="coerce")
                        if pd.isna(month):
                            month = pd.to_datetime(m.group(1), format="%b", errors="coerce")
                        if not pd.isna(month):
                            return _profile(int(pd.Timestamp(month).month), int(m.group(2)), "")
                    except Exception:
                        pass

    ticker_txt = str(ticker or "").strip().upper()
    ticker_profiles = {
        "ANF": _QuarterNarrativeFiscalPeriodProfile(1, 31, "start"),
    }
    return ticker_profiles.get(ticker_txt, _QuarterNarrativeFiscalPeriodProfile())


def _quarter_narrative_resolve_fiscal_period_from_date(
    qd: date,
    profile: _QuarterNarrativeFiscalPeriodProfile,
) -> Tuple[int, int, str, date]:
    candidates = [
        _quarter_narrative_safe_date(int(qd.year) + year_offset, profile.year_end_month, profile.year_end_day)
        for year_offset in (-1, 0, 1)
    ]
    eligible = [cand for cand in candidates if -10 <= (cand - qd).days <= 370]
    fy_end = min(eligible or candidates, key=lambda cand: abs((cand - qd).days))
    days_to_fy_end = (fy_end - qd).days
    if days_to_fy_end <= 45:
        fq = 4
    elif days_to_fy_end <= 135:
        fq = 3
    elif days_to_fy_end <= 225:
        fq = 2
    else:
        fq = 1
    fy = int(fy_end.year) - 1 if profile.year_label == "start" else int(fy_end.year)
    return fy, fq, f"{fy}-Q{fq}", fy_end

@dataclass(frozen=True)
class QuarterNarrativeRecord:
    ticker: str
    fiscal_period: str
    source_period: str = ""
    source_date: str = ""
    source_type: str = ""
    source_file: str = ""
    source_note: str = ""
    category: str = ""
    theme: str = ""
    what_happened: str = ""
    management_framing: str = ""
    why_it_matters: str = ""
    model_implication: str = ""
    valuation_implication: str = ""
    double_count_guardrail: str = ""
    linked_sheet: str = ""
    linked_metric: str = ""
    amount: Any = ""
    unit: str = ""
    confidence: str = "medium"
    include_in_quarter_notes: bool = True
    include_in_promise_progress: bool = False
    include_in_investment_case: bool = False
    raw_quote_short: str = ""
    raw_quote_exact: str = ""


QUARTER_NARRATIVE_DATA_HEADERS = [
    "Ticker",
    "Quarter",
    "Category",
    "Theme",
    "What happened",
    "Management framing",
    "Why it matters",
    "Model implication",
    "Valuation implication",
    "Double-count guardrail",
    "Linked sheet",
    "Linked metric",
    "Amount",
    "Unit",
    "Source date",
    "Source type",
    "Source / note",
    "Confidence",
    "Include in UI",
]


def _quarter_narrative_record_to_audit_row(record: QuarterNarrativeRecord) -> List[Any]:
    source_bits = [record.source_file, record.source_note]
    source_note = " | ".join(str(bit).strip() for bit in source_bits if str(bit or "").strip())
    if record.source_period:
        source_note = f"{source_note} | stated in {record.source_period}" if source_note else f"stated in {record.source_period}"
    return [
        record.ticker,
        record.fiscal_period,
        record.category,
        record.theme,
        record.what_happened,
        record.management_framing,
        record.why_it_matters,
        record.model_implication,
        record.valuation_implication,
        record.double_count_guardrail,
        record.linked_sheet,
        record.linked_metric,
        record.amount,
        record.unit,
        record.source_date,
        record.source_type,
        source_note,
        record.confidence,
        "Yes" if record.include_in_quarter_notes else "No",
    ]


def _quarter_narrative_records_for_ticker(ticker: Any) -> List[QuarterNarrativeRecord]:
    """Conservative structured backing records for Quarter_Notes_UI."""
    ticker_txt = str(ticker or "").strip().upper()
    records: List[QuarterNarrativeRecord] = []

    def _rec(**kwargs: Any) -> None:
        kwargs.setdefault("ticker", ticker_txt)
        records.append(QuarterNarrativeRecord(**kwargs))

    if ticker_txt == "PBI":
        base_source_type = "earnings release / presentation / transcript metadata"
        _rec(
            fiscal_period="2024-Q2",
            source_period="2024-Q2",
            source_date="2024-08-08",
            source_type=base_source_type,
            source_file="PBI/earnings_release/8-K_2024-08-08_earnings_release_q2_2024.htm",
            category="Cost savings / restructuring",
            theme="Implemented cost reductions",
            what_happened="$70m annualized reductions were disclosed as implemented progress.",
            management_framing="Savings came from corporate, SendTech and Presort actions.",
            why_it_matters="It is evidence of run-rate progress, not the full program target.",
            model_implication="Use as achieved/run-rate context for cost savings baseline.",
            valuation_implication="Supports the EBIT bridge only if not double-counted with targets.",
            double_count_guardrail="Do not treat achieved run-rate as the full cost savings target.",
            linked_sheet="Promise_Progress_UI; Investment_Case; Quarter_Notes_UI",
            linked_metric="Cost savings target / run-rate",
            amount=70.0,
            unit="$m annualized",
            confidence="high",
            include_in_promise_progress=True,
            include_in_investment_case=True,
            raw_quote_short="$70m annualized reductions",
        )
        _rec(
            fiscal_period="2024-Q2",
            source_period="2024-Q2",
            source_date="2024-08-08",
            source_type=base_source_type,
            source_file="PBI/earnings_release/8-K_2024-08-08_earnings_release_q2_2024.htm",
            category="Cost savings / restructuring",
            theme="Annual savings target",
            what_happened="$120m-$160m annual savings target was reiterated.",
            management_framing="Management presented the target as a program goal.",
            why_it_matters="The target is a commitment to track against later progress.",
            model_implication="Manual inputs can show the total target, while the bridge uses incremental savings vs baseline.",
            valuation_implication="Only source-backed incremental savings should lift EBITDA/EPS.",
            double_count_guardrail="Target is not achieved savings and should not be added on top of realized run-rate.",
            linked_sheet="Promise_Progress_UI; Investment_Case; Scenario Driver Bridge",
            linked_metric="Incremental cost savings vs baseline",
            amount="$120m-$160m",
            unit="$m annualized target",
            confidence="high",
            include_in_promise_progress=True,
            include_in_investment_case=True,
            raw_quote_short="$120m-$160m annual savings",
        )
        _rec(
            fiscal_period="2024-Q2",
            source_period="2024-Q2",
            source_date="2024-08-08",
            source_type=base_source_type,
            source_file="PBI/earnings_release/8-K_2024-08-08_earnings_release_q2_2024.htm",
            category="Cost savings / restructuring",
            theme="GEC loss removal",
            what_happened="GEC exit was presented as eliminating $136m of annualized net losses.",
            management_framing="Management separated GEC exit economics from operating cost reductions.",
            why_it_matters="Structural loss removal can improve earnings power but is not the same item as cost savings.",
            model_implication="Track separately from cost savings and illustrative EBIT bridge.",
            valuation_implication="May inform normalized EBIT but should not be stacked with the same bridge item twice.",
            double_count_guardrail="Do not double-count GEC loss removal with cost savings or the illustrative EBIT bridge.",
            linked_sheet="Investment_Case; Quarter_Notes_UI",
            linked_metric="GEC loss removal",
            amount=136.0,
            unit="$m annualized loss removal",
            confidence="high",
            include_in_investment_case=True,
            raw_quote_short="$136m annualized net loss removal",
        )
        _rec(
            fiscal_period="2024-Q2",
            source_period="2024-Q2",
            source_date="2024-08-08",
            source_type=base_source_type,
            source_file="PBI/earnings_release/8-K_2024-08-08_earnings_release_q2_2024.htm",
            category="FCF / cash flow",
            theme="Cash optimization",
            what_happened="Go-forward cash needs were framed around $240m after cash optimization.",
            management_framing="Management described cash optimization separately from FCF guidance.",
            why_it_matters="It affects liquidity and refinancing analysis but is not an FCF target.",
            model_implication="Use as cash/liquidity context, not as ordinary FCF actual.",
            valuation_implication="Relevant to debt runway and financing risk.",
            double_count_guardrail="Do not map cash-needs optimization into FCF guidance.",
            linked_sheet="Quarter_Notes_UI; Valuation; Debt_Profile",
            linked_metric="Cash optimization / go-forward cash needs",
            amount=240.0,
            unit="$m go-forward cash needs",
            confidence="high",
            raw_quote_short="$240m go-forward cash needs",
        )
        _rec(
            fiscal_period="2024-Q2",
            source_period="2024-Q2",
            source_date="2024-08-08",
            source_type=base_source_type,
            source_file="PBI/earnings_release/8-K_2024-08-08_earnings_release_q2_2024.htm",
            category="Accounting / non-GAAP definitions",
            theme="Illustrative EBIT bridge",
            what_happened="$481m illustrative EBIT bridge was shown as a non-GAAP bridge, not formal guidance.",
            management_framing="Bridge used TTM EBIT, GEC loss removal and cost-cut midpoint.",
            why_it_matters="It is useful context but not a guidance promise by itself.",
            model_implication="Keep separate from Forward Adj EBIT and cost savings inputs.",
            valuation_implication="Can frame upside case after reconciliation.",
            double_count_guardrail="Do not treat the illustrative EBIT bridge as a forecast or add it to cost savings again.",
            linked_sheet="Quarter_Notes_UI; Investment_Case",
            linked_metric="Illustrative EBIT bridge",
            amount=481.0,
            unit="$m illustrative EBIT",
            confidence="high",
            raw_quote_short="$481m illustrative EBIT bridge",
        )
        _rec(
            fiscal_period="2026-Q1",
            source_period="2026-Q1",
            source_date="2026-05-05",
            source_type="earnings release / presentation",
            source_file="PBI/earnings_release/8-K_2026-05-05_earnings_release_q1_2026.htm",
            category="Cost savings / restructuring",
            theme="Run-rate savings progress",
            what_happened="$157m annualized run-rate savings were disclosed against the current target.",
            management_framing="Management used run-rate progress to show execution toward the savings program.",
            why_it_matters="It is the baseline for incremental cost savings bridge math.",
            model_implication="Manual inputs can show total target/run-rate context; bridge uses target minus baseline.",
            valuation_implication="Incremental savings can affect EBITDA and EPS after tax.",
            double_count_guardrail="Do not backfill later run-rate savings into earlier quarters.",
            linked_sheet="Promise_Progress_UI; Investment_Case; Scenario Driver Bridge",
            linked_metric="Incremental cost savings vs baseline",
            amount=157.0,
            unit="$m annualized run-rate",
            confidence="high",
            include_in_promise_progress=True,
            include_in_investment_case=True,
            raw_quote_short="$157m run-rate",
        )
        _rec(
            fiscal_period="2025-Q4",
            source_period="2025-Q4",
            source_date="2025-12-31",
            source_type="earnings release / guidance profile",
            source_file="PBI/earnings_release/8-K_2026-02-17_earnings_release_q4_2025.htm",
            category="FCF / cash flow",
            theme="FCF definition",
            what_happened="FCF and Adjusted FCF targets are source-specific and should not be collapsed.",
            management_framing="Cash generation guidance is central to debt reduction execution.",
            why_it_matters="Actual/progress must match the source definition used for the target.",
            model_implication="Forward FCF input should use ordinary FCF when the company guides ordinary FCF.",
            valuation_implication="FCF yield and deleveraging analysis depend on the same definition.",
            double_count_guardrail="Do not use adjusted FCF as ordinary FCF, or TTM FCF as annual actual, unless the row says so.",
            linked_sheet="Promise_Progress_UI; Investment_Case; Valuation",
            linked_metric="FCF target / Adjusted FCF target",
            unit="definition",
            confidence="medium",
            include_in_promise_progress=True,
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2026-Q1",
            source_period="2026-Q1",
            source_date="2026-03-31",
            source_type="Promise_Progress_UI / Investment_Case",
            source_file="PBI workbook curated guidance profile",
            category="Debt / liquidity / refinancing",
            theme="Debt and refinancing execution",
            what_happened="Debt, maturities and refinancing remain core diligence items for the turnaround.",
            management_framing="The equity case needs durable cash generation to lower refinancing risk.",
            why_it_matters="FCF credibility matters because balance-sheet risk can dominate the valuation.",
            model_implication="Track debt paydown and interest/refi scenarios separately from operating savings.",
            valuation_implication="Lower net debt and refinancing risk can improve equity value per share.",
            double_count_guardrail="Do not treat debt paydown as EBITDA, and do not invent interest savings without a clean rate/effect.",
            linked_sheet="Promise_Progress_UI; Investment_Case; Valuation; Debt_Profile",
            linked_metric="Debt/refinancing execution",
            unit="watch item",
            confidence="medium",
            include_in_promise_progress=True,
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2026-Q1",
            source_period="2026-Q1",
            source_date="2026-03-31",
            source_type="Promise_Progress_UI / Investment_Case",
            source_file="PBI workbook curated guidance profile",
            category="Segment / brand / geography",
            theme="Presort and SendTech execution",
            what_happened="Presort and SendTech trends remain the recurring earnings-quality proof points.",
            management_framing="Segment stabilization is necessary for savings and FCF to be durable.",
            why_it_matters="Cost actions are more valuable if the segment base is stabilizing.",
            model_implication="Use Segment Scenario Inputs for revenue/margin sensitivity and keep manual stabilization as separate context.",
            valuation_implication="Better segment proof supports a more durable EBIT/FCF multiple.",
            double_count_guardrail="Do not double-enter the same Presort/SendTech uplift through both segment scenario rows and manual bridge rows.",
            linked_sheet="Promise_Progress_UI; Investment_Case; Scenario_Driver_Assumptions",
            linked_metric="Presort / SendTech segment scenario",
            unit="watch item",
            confidence="medium",
            include_in_promise_progress=True,
            include_in_investment_case=True,
        )
    elif ticker_txt == "GPRE":
        _rec(
            fiscal_period="2025-Q4",
            source_period="2025-Q4",
            source_date="2025-12-31",
            source_type="earnings release / operating drivers",
            source_file="GPRE/earnings_release/8-K_2026-02-05_earnings_release_q4_2025.htm",
            category="Policy / regulatory",
            theme="45Z monetization",
            what_happened="$23.4m of 45Z production tax credit value was monetized net of discounts and other costs.",
            management_framing="Management framed 45Z as a material policy-linked earnings and cash driver.",
            why_it_matters="It creates a realized baseline for later 45Z uplift.",
            model_implication="Bridge uses incremental 45Z uplift vs the TTM/reported baseline.",
            valuation_implication="Supports EBITDA/EPS bridge only for incremental uplift.",
            double_count_guardrail="Do not add full 45Z guide on top of TTM if baseline 45Z is already included.",
            linked_sheet="Promise_Progress_UI; Investment_Case; Operating_Drivers; Scenario Driver Bridge",
            linked_metric="Incremental 45Z uplift vs baseline",
            amount=23.4,
            unit="$m",
            confidence="high",
            include_in_promise_progress=True,
            include_in_investment_case=True,
            raw_quote_short="$23.4m in 45Z production tax credit value",
        )
        _rec(
            fiscal_period="2026-Q1",
            source_period="2026-Q1",
            source_date="2026-03-31",
            source_type="earnings release / guidance profile",
            source_file="GPRE/earnings_release/GPRE_Q1_2026_earnings_release.pdf",
            category="Guidance / promise",
            theme="45Z contribution guidance",
            what_happened="2026 year 45Z contribution guidance was stated at $200m-$225m.",
            management_framing="Management excluded on-farm practice upside pending final Treasury guidance/calculator.",
            why_it_matters="The guide is a total contribution view, not an incremental add-on by itself.",
            model_implication="Active guide minus baseline included drives the Scenario Driver Bridge.",
            valuation_implication="Policy upside can move EBITDA/EPS if the incremental amount is source-backed.",
            double_count_guardrail="Use incremental uplift vs baseline, not the full guidance amount.",
            linked_sheet="Promise_Progress_UI; Investment_Case; Scenario Driver Bridge",
            linked_metric="45Z contribution / guide ($m)",
            amount="$200m-$225m",
            unit="$m guide",
            confidence="high",
            include_in_promise_progress=True,
            include_in_investment_case=True,
            raw_quote_short="$200m-$225m",
        )
        _rec(
            fiscal_period="2026-Q1",
            source_period="2026-Q1",
            source_date="2026-03-31",
            source_type="earnings release / operating drivers",
            source_file="GPRE/earnings_release/GPRE_Q1_2026_earnings_release.pdf",
            category="Policy / regulatory",
            theme="45Z facility qualification progress",
            what_happened="Management indicated all eight operating plants qualified or were expected to qualify for 45Z tax credits from Jan. 1; operational/running/monetizing evidence includes Advantage Nebraska operational.",
            management_framing="The company framed qualification as broad facility readiness while final policy/calculator details remain open.",
            why_it_matters="Operational/running/monetizing language is valid evidence for facility qualification unless contradicted.",
            model_implication="Use operational/running/monetizing evidence as qualification progress, not as completed annual 45Z EBITDA guidance.",
            valuation_implication="Facility readiness improves credibility of the 45Z bridge.",
            double_count_guardrail="Facility qualification is progress evidence; do not mark guidance as completed or add full 45Z guidance on top of any 45Z already in baseline.",
            linked_sheet="Promise_Progress_UI; Operating_Drivers; Quarter_Notes_UI",
            linked_metric="45Z facility qualification",
            amount="8 of 8 qualification evidence",
            unit="facilities",
            confidence="high",
            include_in_promise_progress=True,
            raw_quote_short="All eight operating plants qualify/expected to qualify; Advantage Nebraska operational",
        )
        _rec(
            fiscal_period="2025-Q4",
            source_period="2025-Q4",
            source_date="2025-12-31",
            source_type="earnings release / operating drivers",
            source_file="GPRE/earnings_release/8-K_2026-02-05_earnings_release_q4_2025.htm",
            category="Policy / regulatory",
            theme="Advantage Nebraska",
            what_happened="Advantage Nebraska was described as operational.",
            management_framing="Management tied the facility to 45Z contribution potential.",
            why_it_matters="It is facility-specific evidence supporting the broader 45Z qualification pathway.",
            model_implication="Use as progress evidence, not separate unrelated milestone economics.",
            valuation_implication="Improves confidence in 45Z monetization timing.",
            double_count_guardrail="Avoid counting Advantage Nebraska both as facility progress and full 45Z guidance.",
            linked_sheet="Promise_Progress_UI; Operating_Drivers; Quarter_Notes_UI",
            linked_metric="Advantage Nebraska startup",
            amount="AN operational",
            unit="status",
            confidence="high",
            include_in_promise_progress=True,
        )
        _rec(
            fiscal_period="2025-Q4",
            source_period="2025-Q4",
            source_date="2025-12-31",
            source_type="earnings release / guidance profile",
            source_file="GPRE/earnings_release/8-K_2026-02-05_earnings_release_q4_2025.htm",
            category="FCF / cash flow",
            theme="Capex guidance",
            what_happened="2026 sustaining capex guidance was stated at $15m-$25m.",
            management_framing="Capex was framed as a cash-flow item.",
            why_it_matters="Capex affects FCF but not EBITDA or EPS directly.",
            model_implication="Scenario bridge uses capex change vs TTM baseline.",
            valuation_implication="Lower capex can lift FCF after baseline adjustment.",
            double_count_guardrail="Do not subtract total capex again from active FCF.",
            linked_sheet="Investment_Case; Scenario Driver Bridge",
            linked_metric="Capex change vs baseline",
            amount="$15m-$25m",
            unit="$m guide",
            confidence="high",
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2025-Q3",
            source_period="2025-Q3",
            source_date="2025-09-30",
            source_type="earnings release / debt detail",
            source_file="GPRE/earnings_release/8-K_2025-11-05_earnings_release_q3_2025.htm",
            category="Debt / liquidity / refinancing",
            theme="Debt reduction",
            what_happened="Sale proceeds were used to repay junior mezzanine debt.",
            management_framing="Management used the action to reduce balance-sheet risk.",
            why_it_matters="Debt repayment supports liquidity and equity value analysis.",
            model_implication="Track in debt detail and valuation, not as EBITDA uplift.",
            valuation_implication="Lower net debt can improve equity value per share.",
            double_count_guardrail="Do not treat debt repayment as operating EBITDA.",
            linked_sheet="Promise_Progress_UI; Debt_Profile; Valuation",
            linked_metric="Debt reduction",
            amount=130.7,
            unit="$m debt repaid",
            confidence="high",
            include_in_promise_progress=True,
            raw_quote_short="Debt repaid",
        )
        _rec(
            fiscal_period="2025-Q4",
            source_period="2025-Q4",
            source_date="2025-12-31",
            source_type="earnings release / market commentary",
            source_file="GPRE/earnings_release/8-K_2026-02-05_earnings_release_q4_2025.htm",
            category="Commodity / market drivers",
            theme="Crush and policy drivers",
            what_happened="Crush margin and policy/RVO/E15/export rows remain scenario-driven operating reads.",
            management_framing="Commodity spreads are cyclical and policy outcomes can move the case.",
            why_it_matters="They can be material but require explicit numeric scenario inputs.",
            model_implication="Use manual incremental bridge values when a clean dollar effect is entered.",
            valuation_implication="Upside should flow through taxable operating uplift if source-backed.",
            double_count_guardrail="No hidden Economics_Overlay dependency in Scenario Driver Bridge formulas.",
            linked_sheet="Investment_Case; Scenario Driver Bridge; Quarter_Notes_UI",
            linked_metric="Crush margin uplift / Policy uplift",
            unit="scenario input",
            confidence="medium",
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2026-Q1",
            source_period="2026-Q1",
            source_date="2026-03-31",
            source_type="Operating_Drivers / Valuation",
            source_file="GPRE workbook curated operating drivers",
            category="FCF / cash flow",
            theme="FCF and liquidity conversion",
            what_happened="45Z, crush and policy upside still need to convert into cash through capex and working-capital cycles.",
            management_framing="Commodity cycles require disciplined liquidity management.",
            why_it_matters="The investment case improves only if policy/commodity earnings become usable cash flow.",
            model_implication="Use Capex change vs baseline as an FCF-only bridge driver; keep capex and FCF conversion separate from EBITDA add-backs.",
            valuation_implication="Balance-sheet resilience depends on cash conversion, not just reported EBITDA.",
            double_count_guardrail="Do not add capex/working-capital effects to EPS unless a clean earnings link exists.",
            linked_sheet="Operating_Drivers; Investment_Case; Valuation",
            linked_metric="Capex change vs baseline / FCF conversion",
            unit="watch item",
            confidence="medium",
            include_in_investment_case=True,
        )
    elif ticker_txt == "ANF":
        source_type = "earnings release / guidance profile"
        _rec(
            fiscal_period="2025-Q4",
            source_period="2025-Q4",
            source_date="2026-03-04",
            source_type=source_type,
            source_file="ANF/earnings_release/8-K_2026-03-04_earnings_release.htm",
            category="Guidance / promise",
            theme="Sales growth and brand/geography reads",
            what_happened="2025 net sales growth, Abercrombie/Hollister brand reads and Americas/EMEA/APAC geography reads remain central to the demand story.",
            management_framing="Hollister momentum, Abercrombie stabilization and Americas/EMEA/APAC store reads frame the growth quality.",
            why_it_matters="The revenue story is stronger when brand and geography cuts agree with total sales growth.",
            model_implication="Use segment rows as separate brand or geography bases; do not sum both cuts.",
            valuation_implication="Sustained sales growth supports the EPS/FCF multiple if margins hold.",
            double_count_guardrail="Do not double-count brand and geography revenue assumptions.",
            linked_sheet="Promise_Progress_UI; Operating_Drivers; Investment_Case; Scenario_Driver_Assumptions",
            linked_metric="Sales growth / Segment Scenario Inputs",
            unit="growth read",
            confidence="high",
            include_in_promise_progress=True,
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2025-Q4",
            source_period="2025-Q4",
            source_date="2026-03-04",
            source_type=source_type,
            source_file="ANF/earnings_release/8-K_2026-03-04_earnings_release.htm",
            category="Earnings / margin",
            theme="Operating margin normalization",
            what_happened="2025 operating margin was evaluated against annual guides while 2026 guide implies a lower margin baseline.",
            management_framing="Management framed tariff, ERP, freight and marketing as important 2026 margin bridge components.",
            why_it_matters="The debate is not just sales growth; margin normalization drives EPS sensitivity.",
            model_implication="Use operating margin guide for baseline context and bps bridge rows for incremental effects.",
            valuation_implication="Lower margin guide can pressure EPS/valuation unless offsets are proven.",
            double_count_guardrail="Do not count the same margin pressure in both operating margin guidance and the bps margin bridge.",
            linked_sheet="Promise_Progress_UI; Investment_Case; Scenario Driver Bridge",
            linked_metric="Operating margin / Margin bridge vs baseline",
            unit="margin read",
            confidence="high",
            include_in_promise_progress=True,
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2026-Q1",
            source_period="2026-Q1",
            source_date="2026-03-04",
            source_type=source_type,
            source_file="ANF/earnings_release/8-K_2026-03-04_earnings_release.htm",
            category="Earnings / margin",
            theme="Tariff headwind",
            what_happened="Tariffs were disclosed as a margin headwind that can be modeled in bps.",
            management_framing="Management expected tariff pressure to be partly offset by other levers.",
            why_it_matters="Bps margin effects can be translated into dollar earnings impact using active revenue.",
            model_implication="Convert tariff bps to $m using active revenue and feed the margin bridge.",
            valuation_implication="Taxable operating drag affects EBITDA and EPS after tax.",
            double_count_guardrail="Do not also add the same tariff effect through operating margin guidance.",
            linked_sheet="Investment_Case; Scenario Driver Bridge; Quarter_Notes_UI",
            linked_metric="Tariff impact (bps)",
            amount="-290 / -70",
            unit="bps",
            confidence="high",
            include_in_investment_case=True,
            raw_quote_short="tariff headwind",
        )
        _rec(
            fiscal_period="2026-Q1",
            source_period="2026-Q1",
            source_date="2026-03-04",
            source_type=source_type,
            source_file="ANF/earnings_release/8-K_2026-03-04_earnings_release.htm",
            category="Earnings / margin",
            theme="Freight tailwind",
            what_happened="Freight was identified as a positive margin offset.",
            management_framing="Freight relief helps offset tariff, ERP and marketing headwinds.",
            why_it_matters="Tailwinds should preserve positive sign in the margin bridge.",
            model_implication="Convert freight bps to $m using active revenue.",
            valuation_implication="Taxable operating uplift can lift bridge-adjusted EBITDA/EPS.",
            double_count_guardrail="Do not net freight into tariff if individual rows are modeled.",
            linked_sheet="Investment_Case; Scenario Driver Bridge; Quarter_Notes_UI",
            linked_metric="Freight tailwind (bps)",
            amount=160,
            unit="bps",
            confidence="high",
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2026-Q1",
            source_period="2026-Q1",
            source_date="2026-03-04",
            source_type=source_type,
            source_file="ANF/earnings_release/8-K_2026-03-04_earnings_release.htm",
            category="Earnings / margin",
            theme="ERP disruption",
            what_happened="ERP transition costs/disruption were framed as a margin headwind.",
            management_framing="Management treated ERP as a temporary operating drag.",
            why_it_matters="ERP headwind belongs in the margin bridge, not in sales growth.",
            model_implication="Convert ERP bps to $m when bps and revenue are available.",
            valuation_implication="Taxable operating drag reduces bridge-adjusted EBITDA/EPS.",
            double_count_guardrail="Do not also include the same ERP drag in a separate manual EBITDA row.",
            linked_sheet="Investment_Case; Scenario Driver Bridge; Quarter_Notes_UI",
            linked_metric="ERP disruption (bps)",
            amount="-100+",
            unit="bps",
            confidence="medium",
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2026-Q1",
            source_period="2026-Q1",
            source_date="2026-03-04",
            source_type=source_type,
            source_file="ANF/earnings_release/8-K_2026-03-04_earnings_release.htm",
            category="Earnings / margin",
            theme="Marketing headwind",
            what_happened="Marketing was identified as a margin headwind.",
            management_framing="Marketing spend supports brand demand but weighs on margin.",
            why_it_matters="It is an operating margin bridge item with clear bps treatment.",
            model_implication="Convert marketing bps to $m using active revenue.",
            valuation_implication="Taxable operating drag affects bridge-adjusted EBITDA/EPS.",
            double_count_guardrail="Do not double-count marketing through both bps bridge and operating margin guide.",
            linked_sheet="Investment_Case; Scenario Driver Bridge; Quarter_Notes_UI",
            linked_metric="Marketing headwind (bps)",
            amount=-50,
            unit="bps",
            confidence="high",
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2025-Q4",
            source_period="2025-Q4",
            source_date="2026-03-04",
            source_type="earnings release / capital allocation",
            source_file="ANF/earnings_release/8-K_2026-03-04_earnings_release.htm",
            category="Capital allocation / buybacks",
            theme="Buybacks and share count",
            what_happened="2025 buybacks were about $450m and reduced diluted share count context.",
            management_framing="Capital return remained active while net cash stayed strong.",
            why_it_matters="Buybacks affect EPS through diluted shares, not EBITDA.",
            model_implication="Active diluted shares drive Bridge EPS denominator.",
            valuation_implication="Per-share valuation changes through share count and cash use.",
            double_count_guardrail="Do not add a separate buyback EPS benefit if active shares already reflect buybacks.",
            linked_sheet="Investment_Case; Scenario Driver Bridge; Valuation",
            linked_metric="Buyback amount / Diluted shares",
            amount=450.0,
            unit="$m buybacks",
            confidence="high",
            include_in_promise_progress=True,
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2025-Q4",
            source_period="2025-Q4",
            source_date="2026-03-04",
            source_type="earnings release / guidance profile",
            source_file="ANF/earnings_release/8-K_2026-03-04_earnings_release.htm",
            category="Accounting / non-GAAP definitions",
            theme="EPS basis",
            what_happened="Adjusted EPS and GAAP EPS should be tracked with source-specific labels.",
            management_framing="Management shows adjusted and GAAP basis separately where relevant.",
            why_it_matters="Generic EPS labels can compare guidance and actuals on the wrong basis.",
            model_implication="Use direct_eps only for explicit EPS-per-share impacts.",
            valuation_implication="P/E read should use meaningful EPS basis and N/M when not meaningful.",
            double_count_guardrail="Do not merge Adjusted EPS and GAAP EPS unless the source does.",
            linked_sheet="Promise_Progress_UI; Investment_Case; Scenario_Bridge_Tax_Treatment",
            linked_metric="Adjusted EPS / GAAP EPS basis",
            unit="definition",
            confidence="high",
            include_in_promise_progress=True,
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2025-Q4",
            source_period="Jan 2026 pre-release update",
            source_date="2026-01-12",
            source_type="pre-release update",
            source_file="ANF/press_release/ANF_2026-01-12_press_release_business_update.pdf",
            category="Guidance / promise",
            theme="Pre-release update",
            what_happened="January pre-release narrowed 2025 guidance before the final Q4 release.",
            management_framing="Management updated the annual view before final results.",
            why_it_matters="It is a separate event from the final Q4 actual.",
            model_implication="Keep pre-release guide as an update event; final Q4 fills Actual for the same annual horizon.",
            valuation_implication="Tightened guide improves tracking but should not replace final actuals.",
            double_count_guardrail="Do not create separate final-actual rows when the Actual cell can be filled.",
            linked_sheet="Promise_Progress_UI; Quarter_Notes_UI",
            linked_metric="2025-Q4 pre-release update",
            unit="event",
            confidence="high",
            include_in_promise_progress=True,
        )
        _rec(
            fiscal_period="2025-Q4",
            source_period="2025-Q4",
            source_date="2026-03-04",
            source_type=source_type,
            source_file="ANF/earnings_release/8-K_2026-03-04_earnings_release.htm",
            category="FCF / cash flow",
            theme="Capex discipline",
            what_happened="Capex was tracked against annual guidance and remains a cash-flow driver.",
            management_framing="Store activity and investment needs sit alongside buybacks and FCF.",
            why_it_matters="Capex affects FCF and cash available for repurchases, but not EBITDA directly.",
            model_implication="Treat capex as change vs baseline in FCF, not an EBITDA/EPS item.",
            valuation_implication="FCF yield and buyback capacity depend on capex discipline.",
            double_count_guardrail="Do not subtract total capex again if active FCF already embeds baseline capex.",
            linked_sheet="Promise_Progress_UI; Investment_Case; Valuation",
            linked_metric="Capex",
            unit="$m capex",
            confidence="high",
            include_in_promise_progress=True,
            include_in_investment_case=True,
        )
        _rec(
            fiscal_period="2026-Q1",
            source_period="2026-Q1",
            source_date="2026-03-04",
            source_type=source_type,
            source_file="ANF/earnings_release/8-K_2026-03-04_earnings_release.htm",
            category="Segment / brand / geography",
            theme="Brand and geography cuts",
            what_happened="Brand rows and geography/store rows are separate cuts of company revenue.",
            management_framing="Abercrombie/Hollister and Americas/EMEA/APAC describe different views.",
            why_it_matters="Both cuts can inform the story, but they cannot be summed together.",
            model_implication="Active basis controls whether Brand, Geography or None feeds the segment bridge.",
            valuation_implication="Segment scenario impact is taxable operating uplift only for selected basis.",
            double_count_guardrail="Never sum brand and geography rows together.",
            linked_sheet="Investment_Case; Scenario_Driver_Assumptions; Quarter_Notes_UI",
            linked_metric="Segment Scenario Inputs",
            unit="basis guardrail",
            confidence="high",
            include_in_investment_case=True,
        )

    return records


def _write_quarter_narrative_data_sheet(
    wb: Workbook,
    ticker: Any,
    records: Optional[Sequence[QuarterNarrativeRecord]] = None,
) -> None:
    if "Quarter_Narrative_Data" in wb.sheetnames:
        del wb["Quarter_Narrative_Data"]
    ws = wb.create_sheet("Quarter_Narrative_Data")
    records_list = list(records if records is not None else _quarter_narrative_records_for_ticker(ticker))
    ws.append(QUARTER_NARRATIVE_DATA_HEADERS)
    for record in records_list:
        ws.append(_quarter_narrative_record_to_audit_row(record))

    header_fill = PatternFill("solid", fgColor="5B9BD5")
    header_font = Font(bold=True, color="FFFFFF")
    body_fill = PatternFill("solid", fgColor="F7FBFF")
    alt_fill = PatternFill("solid", fgColor="EDF4FB")
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for cell in ws[1]:
        cell.fill = copy(header_fill)
        cell.font = copy(header_font)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = copy(thin_border)
    narrative_cols = {5, 6, 7, 8, 9, 10, 17}
    for rr in range(2, int(ws.max_row or 1) + 1):
        fill = alt_fill if rr % 2 == 0 else body_fill
        for cc in range(1, len(QUARTER_NARRATIVE_DATA_HEADERS) + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = copy(fill)
            cell.border = copy(thin_border)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=cc in narrative_cols)
        ws.row_dimensions[rr].height = 42.0
    widths = {
        "A": 10,
        "B": 12,
        "C": 24,
        "D": 26,
        "E": 42,
        "F": 34,
        "G": 36,
        "H": 38,
        "I": 34,
        "J": 38,
        "K": 32,
        "L": 28,
        "M": 18,
        "N": 16,
        "O": 14,
        "P": 26,
        "Q": 42,
        "R": 12,
        "S": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:S{max(int(ws.max_row or 1), 1)}"


def _quarter_narrative_period_sort_key(period: Any) -> Tuple[int, int]:
    txt = str(period or "").strip()
    match = re.fullmatch(r"(20\d{2})-Q([1-4])", txt, flags=re.I)
    if match:
        return (int(match.group(1)), int(match.group(2)))
    year_match = re.search(r"(20\d{2})", txt)
    return (int(year_match.group(1)) if year_match else 0, 0)


def _quarter_narrative_source_label(record: QuarterNarrativeRecord) -> str:
    parts = [record.confidence.title() if record.confidence else "", record.source_type, record.source_date]
    return "; ".join(str(part).strip() for part in parts if str(part or "").strip())


def _quarter_narrative_compact_sentence(text: Any, *, max_chars: int = 185) -> str:
    txt = re.sub(r"\s+", " ", str(text or "").strip())
    if len(txt) <= max_chars:
        return txt
    cut = txt[: max(1, int(max_chars) - 1)].rsplit(" ", 1)[0].rstrip(" ,;:")
    return f"{cut}."


def _quarter_narrative_row_height(*texts: Any, base: float = 34.0, max_height: float = 90.0) -> float:
    longest = max((len(str(text or "")) for text in texts), default=0)
    if longest <= 85:
        return base
    if longest <= 160:
        return min(max_height, base + 14.0)
    if longest <= 260:
        return min(max_height, base + 28.0)
    return max_height


def _quarter_narrative_recent_history_periods(wb: Workbook, *, limit: int = 8) -> List[str]:
    """Return newest fiscal quarter labels available in History_Q.

    Quarter_Notes_UI should make sparse recent quarters explicit instead of
    silently jumping over them.  This helper intentionally reads only resolved
    fiscal labels/year+quarter columns so it works for retail fiscal calendars
    as well as calendar-year reporters.
    """
    if "History_Q" not in getattr(wb, "sheetnames", []):
        return []
    ws = wb["History_Q"]
    headers = [str(ws.cell(1, cc).value or "").strip().lower() for cc in range(1, int(ws.max_column or 0) + 1)]
    label_col = None
    for name in ("fiscal_period", "fiscal label", "fiscal_label", "quarter", "period"):
        if name in headers:
            label_col = headers.index(name) + 1
            break
    fy_col = headers.index("fiscal_year") + 1 if "fiscal_year" in headers else None
    fq_col = headers.index("fiscal_quarter") + 1 if "fiscal_quarter" in headers else None

    labels: Set[str] = set()
    for rr in range(2, int(ws.max_row or 0) + 1):
        label = ""
        if label_col:
            raw = str(ws.cell(rr, label_col).value or "").strip()
            match = re.search(r"(20\d{2})-Q([1-4])", raw, flags=re.I)
            if match:
                label = f"{match.group(1)}-Q{match.group(2)}"
        if not label and fy_col and fq_col:
            year_match = re.search(r"20\d{2}", str(ws.cell(rr, fy_col).value or ""))
            q_match = re.search(r"[1-4]", str(ws.cell(rr, fq_col).value or ""))
            if year_match and q_match:
                label = f"{year_match.group(0)}-Q{q_match.group(0)}"
        if label:
            labels.add(label)
    return sorted(labels, key=_quarter_narrative_period_sort_key, reverse=True)[: max(0, int(limit or 0))]


def _quarter_narrative_recent_periods_from_frame(hist: Any, *, ticker: Any = "", limit: int = 8) -> List[str]:
    if not isinstance(hist, pd.DataFrame) or hist.empty:
        return []
    cols = {str(col).strip().lower(): col for col in hist.columns}
    label_col = next(
        (cols[name] for name in ("fiscal_period", "fiscal label", "fiscal_label", "quarter", "period") if name in cols),
        None,
    )
    fy_col = cols.get("fiscal_year")
    fq_col = cols.get("fiscal_quarter")
    labels: Set[str] = set()
    profile = _quarter_narrative_fiscal_profile_from_workbook(None, ticker)
    for _, row in hist.iterrows():
        label = ""
        if label_col is not None:
            raw_value = row.get(label_col, "")
            match = re.search(r"(20\d{2})-Q([1-4])", str(raw_value or ""), flags=re.I)
            if match:
                label = f"{match.group(1)}-Q{match.group(2)}"
            else:
                q_ts = pd.to_datetime(raw_value, errors="coerce")
                if pd.notna(q_ts):
                    _fy, _fq, resolved_label, _fy_end = _quarter_narrative_resolve_fiscal_period_from_date(pd.Timestamp(q_ts).date(), profile)
                    label = resolved_label
        if not label and fy_col is not None and fq_col is not None:
            year_match = re.search(r"20\d{2}", str(row.get(fy_col, "") or ""))
            q_match = re.search(r"[1-4]", str(row.get(fq_col, "") or ""))
            if year_match and q_match:
                label = f"{year_match.group(0)}-Q{q_match.group(0)}"
        if label:
            labels.add(label)
    return sorted(labels, key=_quarter_narrative_period_sort_key, reverse=True)[: max(0, int(limit or 0))]


_QUARTER_NARRATIVE_NOISE_PATTERNS = (
    "[updated]",
    "debug",
    "todo",
    "fixme",
    "raw_json",
    "metadata_candidate",
    "source_txt_file",
    "source_txt",
    "<html",
    "<?xml",
    "nan",
    " none ",
    " null ",
)


def _quarter_narrative_clean_text(value: Any, *, max_chars: int = 260) -> str:
    txt = re.sub(r"\s+", " ", str(value or "").strip())
    if not txt:
        return ""
    low = f" {txt.lower()} "
    if any(pattern in low for pattern in _QUARTER_NARRATIVE_NOISE_PATTERNS):
        return ""
    txt = ILLEGAL_CHARACTERS_RE.sub("", txt)
    if len(txt) <= max_chars:
        return txt
    return _quarter_narrative_compact_sentence(txt, max_chars=max_chars)


def _quarter_narrative_period_from_source_quarter(value: Any, *, ticker: Any = "") -> str:
    q_ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(q_ts):
        raw = str(value or "").strip()
        match = re.search(r"(20\d{2})-Q([1-4])", raw, flags=re.I)
        if match:
            return f"{match.group(1)}-Q{match.group(2)}"
        return ""
    profile = _quarter_narrative_fiscal_profile_from_workbook(None, ticker)
    _fy, _fq, label, _fy_end = _quarter_narrative_resolve_fiscal_period_from_date(pd.Timestamp(q_ts).date(), profile)
    return str(label or "").strip()


def _quarter_narrative_category_theme(row: Mapping[str, Any]) -> Tuple[str, str]:
    topic = str(row.get("topic") or row.get("tag") or "").strip()
    category = str(row.get("category") or "").strip()
    headline = str(row.get("headline") or row.get("claim") or topic or "Quarter driver").strip()
    blob = " ".join(str(row.get(key) or "") for key in ("topic", "tag", "category", "headline", "claim", "metric_ref")).lower()
    if any(term in blob for term in ("guidance", "target", "promise", "outlook")):
        return "Guidance / promise", headline
    if any(term in blob for term in ("45z", "rvo", "rin", "e15", "policy", "treasury", "facility", "advantage nebraska")):
        return "Policy / regulatory", headline
    if any(term in blob for term in ("fcf", "cash flow", "liquidity", "cash conversion")):
        return "FCF / cash flow", headline
    if any(term in blob for term in ("debt", "revolver", "refi", "leverage", "covenant")):
        return "Debt / liquidity / refinancing", headline
    if any(term in blob for term in ("buyback", "share count", "shares", "capital allocation", "equity")):
        return "Capital allocation / buybacks", headline
    if any(term in blob for term in ("brand", "geography", "segment", "presort", "sendtech", "abercrombie", "hollister")):
        return "Segment / brand / geography", headline
    if any(term in blob for term in ("cost", "restructuring", "savings", "gec")):
        return "Cost savings / restructuring", headline
    if any(term in blob for term in ("margin", "ebit", "ebitda", "eps", "revenue", "sales")):
        return "Earnings / margin", headline
    if category:
        return category, headline
    return "Quarter narrative", headline


def _quarter_narrative_implications_for_row(ticker: Any, row: Mapping[str, Any], category: str, theme: str) -> Tuple[str, str, str, str]:
    ticker_txt = str(ticker or "").strip().upper()
    blob = " ".join(
        str(row.get(key) or "")
        for key in ("topic", "tag", "category", "headline", "claim", "body", "render_summary", "metric_ref", theme)
    ).lower()
    why = "Source-backed quarter item helps explain what moved and what needs follow-up."
    model = "Use as context unless it maps to a clean Investment_Case or operating-driver input."
    valuation = "Use for narrative support around the quarter's valuation drivers."
    guardrail = "Do not turn commentary into numeric guidance unless the source gives a clean metric."
    if "debt" in category.lower() or "revolver" in blob or "liquidity" in blob:
        why = "Balance-sheet capacity and refinancing/liquidity changes can alter risk and equity value."
        model = "Tie to debt, cash and refinancing assumptions; avoid hidden capital-structure assumptions."
        valuation = "Affects net debt/cash, risk read and multiple support."
        guardrail = "Do not double-count debt changes in both net debt and scenario EPS unless explicitly modeled."
    elif "fcf" in blob or "cash flow" in blob:
        why = "Cash conversion determines whether earnings translate into value and debt capacity."
        model = "Use only definition-compatible FCF/Adjusted FCF in Valuation, Promise and Investment_Case."
        valuation = "Supports FCF-yield and balance-sheet valuation reads."
        guardrail = "Do not mix ordinary FCF, adjusted FCF, CFO or TTM values without a visible basis."
    elif "margin" in blob or "ebit" in blob or "ebitda" in blob:
        why = "Margin conversion shows whether revenue changes are flowing through to earnings."
        model = "Map to operating margin/EBIT/EBITDA drivers only when the definition is clean."
        valuation = "Supports EBITDA/EPS bridge and multiple confidence."
        guardrail = "Do not stack margin bridge bps on top of operating margin guidance unless incremental."
    elif "revenue" in blob or "sales" in blob:
        why = "Revenue/sales trends set the baseline for margin, EPS and segment scenarios."
        model = "Use as revenue context; keep quarterly, YTD and annual horizons distinct."
        valuation = "Revenue quality affects growth/multiple support."
        guardrail = "Do not use TTM or annual values as quarter actuals."
    elif "buyback" in blob or "share count" in blob or "shares" in blob:
        why = "Share-count movement changes EPS denominator and capital-allocation quality."
        model = "Feed EPS through diluted shares; do not affect EBITDA."
        valuation = "Supports EPS/share value and capital allocation read."
        guardrail = "Do not let buybacks increase EBITDA or cash flow unless explicitly modeled."
    elif "cost" in blob or "savings" in blob or "restructuring" in blob or "gec" in blob:
        why = "Cost actions can improve run-rate earnings but often overlap with restructuring or exit benefits."
        model = "Separate target, achieved/run-rate savings and one-time restructuring costs."
        valuation = "Supports normalized EBIT/EBITDA if the run-rate is source-backed."
        if ticker_txt == "PBI":
            guardrail = "Do not add cost savings, GEC loss removal and illustrative EBIT bridge as fully additive unless sourced."
        elif ticker_txt == "GPRE":
            guardrail = "Do not add cost savings or restructuring benefits on top of 45Z/crush upside unless the source clearly separates them."
        else:
            guardrail = "Do not double-count cost savings or restructuring benefits with normalized margin guidance."
    elif "45z" in blob or "facility" in blob or "rvo" in blob or "policy" in blob:
        why = "Policy and facility progress can change the earnings baseline and scenario upside."
        model = "Use 45Z or policy effects as incremental vs baseline when they feed the bridge."
        valuation = "Supports scenario upside, but timing and baseline overlap matter."
        guardrail = "Do not add a full 45Z guide on top of TTM if partial credits are already in baseline."
    if ticker_txt == "ANF" and any(term in blob for term in ("tariff", "freight", "erp", "marketing")):
        why = "Margin bridge items explain near-term operating-margin pressure or relief."
        model = "Convert bps margin effects to dollars using active revenue when used in Investment_Case."
        valuation = "Affects bridge-adjusted EBITDA/EPS."
        guardrail = "Do not mix bps and dollar impacts without conversion, and avoid double-counting with margin guidance."
    if ticker_txt == "ANF" and any(term in blob for term in ("brand", "geography", "abercrombie", "hollister", "americas", "emea", "apac")):
        guardrail = "Brand and geography rows are alternative cuts; never sum both bases together."
    return why, model, valuation, guardrail


def _quarter_narrative_records_from_quarter_notes(
    ticker: Any,
    quarter_notes: Any,
    *,
    history_periods: Optional[Sequence[str]] = None,
    max_per_period: int = 5,
) -> List[QuarterNarrativeRecord]:
    """Promote clean Quarter_Notes rows into narrative records.

    The legacy Quarter_Notes pipeline already contains source-backed model,
    filing and guidance observations.  The narrative UI should use those clean
    rows instead of falling back to a tiny static list, while still filtering
    noisy raw snippets and preserving the audit trail.
    """
    if not isinstance(quarter_notes, pd.DataFrame) or quarter_notes.empty:
        return []
    qn = quarter_notes.copy()
    if "quarter" not in qn.columns:
        return []
    qn["_period_label"] = qn["quarter"].map(lambda val: _quarter_narrative_period_from_source_quarter(val, ticker=ticker))
    qn = qn[qn["_period_label"].astype(str).str.len() > 0].copy()
    if qn.empty:
        return []
    allowed_periods = {str(period or "").strip() for period in (history_periods or []) if str(period or "").strip()}
    if allowed_periods:
        qn = qn[qn["_period_label"].isin(allowed_periods)].copy()
    if qn.empty:
        return []

    def _boolish(value: Any) -> bool:
        if isinstance(value, bool):
            return bool(value)
        txt = str(value or "").strip().lower()
        return txt in {"true", "1", "yes", "y"}

    def _text_from_row(row: Mapping[str, Any]) -> str:
        for key in ("render_summary", "body", "headline", "claim", "note"):
            txt = _quarter_narrative_clean_text(row.get(key), max_chars=260)
            if txt and len(txt) >= 18:
                return txt
        return ""

    def _score(row: Mapping[str, Any]) -> float:
        score = pd.to_numeric(row.get("render_score"), errors="coerce")
        if pd.isna(score):
            score = pd.to_numeric(row.get("severity_score"), errors="coerce")
        score_f = float(score) if pd.notna(score) else 0.0
        if _boolish(row.get("renderable_note")):
            score_f += 20.0
        confidence = str(row.get("confidence") or "").strip().lower()
        if confidence == "high":
            score_f += 8.0
        elif confidence == "low":
            score_f -= 10.0
        method = str(row.get("method") or "").strip().lower()
        if method in {"keyword_scan"}:
            score_f -= 18.0
        if str(row.get("render_drop_reason") or "").strip():
            score_f -= 60.0
        return score_f

    rows_by_period: Dict[str, List[Tuple[float, Mapping[str, Any]]]] = {}
    for row in qn.to_dict("records"):
        period = str(row.get("_period_label") or "").strip()
        if not period:
            continue
        summary = _text_from_row(row)
        if not summary:
            continue
        if not _boolish(row.get("renderable_note")) and _score(row) < 30.0:
            continue
        rows_by_period.setdefault(period, []).append((_score(row), row))

    records: List[QuarterNarrativeRecord] = []
    ticker_txt = str(ticker or "").strip().upper()
    for period in sorted(rows_by_period, key=_quarter_narrative_period_sort_key, reverse=True):
        seen: Set[Tuple[str, str]] = set()
        added = 0
        for _score_val, row in sorted(rows_by_period[period], key=lambda item: item[0], reverse=True):
            if added >= max(1, int(max_per_period or 1)):
                break
            what = _text_from_row(row)
            if not what:
                continue
            category, theme = _quarter_narrative_category_theme(row)
            theme = _quarter_narrative_clean_text(theme, max_chars=80) or "Quarter driver"
            metric_probe = " ".join(
                str(row.get(key) or "")
                for key in ("metric_ref", "topic", "category", "headline", "claim")
            ).lower()
            if ticker_txt == "ANF" and "margin_bridge" in metric_probe:
                category = "Earnings / margin"
                theme = "Margin bridge"
            dedupe_key = (theme.lower(), what.lower())
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            why, model, valuation, guardrail = _quarter_narrative_implications_for_row(ticker_txt, row, category, theme)
            metric_ref = _quarter_narrative_clean_text(row.get("metric_ref") or row.get("topic") or theme, max_chars=80)
            source_type = _quarter_narrative_clean_text(
                row.get("doc_type") or row.get("method") or row.get("source_type") or "Quarter_Notes",
                max_chars=45,
            )
            if source_type and "source-backed" not in source_type.lower():
                source_type = f"source-backed {source_type}"
            confidence_txt = str(row.get("confidence") or "").strip().lower()
            if confidence_txt == "med":
                confidence_txt = "medium"
            if confidence_txt not in {"high", "medium", "low"}:
                confidence_txt = "medium"
            source_note = _quarter_narrative_clean_text(
                row.get("render_preferred_source") or row.get("evidence_doc") or row.get("source_doc") or row.get("doc") or "",
                max_chars=90,
            )
            source_date = ""
            q_ts = pd.to_datetime(row.get("quarter"), errors="coerce")
            if pd.notna(q_ts):
                source_date = pd.Timestamp(q_ts).date().isoformat()
            records.append(
                QuarterNarrativeRecord(
                    ticker=ticker_txt,
                    fiscal_period=period,
                    source_period=period,
                    source_date=source_date,
                    source_type=source_type or "Quarter_Notes",
                    source_file="",
                    source_note=source_note,
                    category=category,
                    theme=theme,
                    what_happened=what,
                    management_framing=_quarter_narrative_clean_text(row.get("claim") or row.get("headline"), max_chars=150),
                    why_it_matters=why,
                    model_implication=model,
                    valuation_implication=valuation,
                    double_count_guardrail=guardrail,
                    linked_sheet="Quarter_Notes_UI; Operating_Drivers",
                    linked_metric=metric_ref or theme,
                    amount=row.get("metric_value") if pd.notna(pd.to_numeric(row.get("metric_value"), errors="coerce")) else "",
                    unit="",
                    confidence=confidence_txt,
                    include_in_quarter_notes=True,
                    include_in_promise_progress="guidance" in category.lower() or "promise" in category.lower(),
                    include_in_investment_case=any(
                        token in f"{category} {theme} {metric_ref}".lower()
                        for token in ("guidance", "margin", "fcf", "debt", "45z", "cost", "buyback", "segment", "revenue")
                    ),
                    raw_quote_short="",
                    raw_quote_exact="",
                )
            )
            added += 1
    return records


def _quarter_narrative_period_from_label_or_date(value: Any, *, ticker: Any = "") -> str:
    raw = str(value or "").strip()
    match = re.search(r"(20\d{2})-Q([1-4])", raw, flags=re.I)
    if match:
        return f"{match.group(1)}-Q{match.group(2)}"
    return _quarter_narrative_period_from_source_quarter(value, ticker=ticker)


def _quarter_narrative_format_surface_value(value: Any, *, label: Any = "", raw_dollars: bool = False) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        txt = _quarter_narrative_clean_text(value, max_chars=80)
        return txt
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return _quarter_narrative_clean_text(value, max_chars=80)
    num_f = float(num)
    label_low = str(label or "").lower()
    if raw_dollars and any(term in label_low for term in ("revenue", "income", "ebit", "ebitda", "fcf", "cash", "capex", "debt", "cfo")):
        num_f = num_f / 1_000_000.0
    if "%" in label_low or "margin" in label_low or "rate" in label_low:
        if abs(num_f) <= 1.5:
            num_f *= 100.0
        return f"{num_f:.1f}%"
    if any(term in label_low for term in ("eps", "per share")):
        return f"${num_f:.2f}"
    if any(term in label_low for term in ("revenue", "income", "ebit", "ebitda", "fcf", "cash", "capex", "debt", "cfo", "buyback", "45z", "crush")):
        return f"${num_f:,.1f}m"
    if abs(num_f - round(num_f)) < 1e-6:
        return f"{num_f:,.0f}"
    return f"{num_f:,.1f}"


def _quarter_narrative_amount_from_surface_value(value: Any) -> Any:
    """Return a value-like Amount, never descriptor/classification prose."""
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    num = pd.to_numeric(value, errors="coerce")
    if pd.notna(num):
        return value
    txt = _quarter_narrative_clean_text(value, max_chars=160)
    if not txt:
        return ""
    low = txt.lower()
    if any(
        phrase in low
        for phrase in (
            "are geographic segments",
            "are brand",
            "brand families",
            "not additive",
        )
    ):
        return ""
    if not re.search(r"[$%0-9]", txt):
        return ""
    if len(txt) > 60 and not re.search(r"\b(ytd|fy|q[1-4]|run[- ]rate|operational|qualified|of\s+\d+)\b", low):
        return ""
    return txt


def _quarter_narrative_source_date_from_period(period: Any) -> str:
    match = re.fullmatch(r"(20\d{2})-Q([1-4])", str(period or "").strip(), flags=re.I)
    if not match:
        return ""
    year = int(match.group(1))
    quarter = int(match.group(2))
    month = quarter * 3
    if month == 12:
        return f"{year}-12-31"
    return (date(year, month + 1, 1) - timedelta(days=1)).isoformat()


def _quarter_narrative_surface_row_terms(ticker: Any) -> Tuple[str, ...]:
    ticker_txt = str(ticker or "").strip().upper()
    common = (
        "revenue",
        "sales",
        "margin",
        "ebit",
        "ebitda",
        "eps",
        "fcf",
        "cash",
        "capex",
        "debt",
        "liquidity",
        "buyback",
        "shares",
    )
    if ticker_txt == "PBI":
        return common + ("presort", "sendtech", "cost", "savings", "gec", "restructuring")
    if ticker_txt == "GPRE":
        return common + ("45z", "crush", "utilization", "ethanol", "corn", "coproduct", "rin", "rvo", "e15", "advantage")
    if ticker_txt == "ANF":
        return common + (
            "abercrombie",
            "hollister",
            "americas",
            "emea",
            "apac",
            "comp",
            "inventory",
            "tariff",
            "freight",
            "erp",
            "marketing",
        )
    return common


def _quarter_narrative_records_from_history_q(
    wb: Workbook,
    ticker: Any,
    *,
    history_periods: Optional[Sequence[str]] = None,
) -> List[QuarterNarrativeRecord]:
    if "History_Q" not in getattr(wb, "sheetnames", []):
        return []
    ws = wb["History_Q"]
    headers = [str(ws.cell(1, cc).value or "").strip().lower() for cc in range(1, int(ws.max_column or 0) + 1)]
    header_map = {header: idx + 1 for idx, header in enumerate(headers) if header}
    label_col = next((header_map[name] for name in ("fiscal_period", "fiscal label", "fiscal_label", "quarter", "period") if name in header_map), None)
    if not label_col:
        return []
    allowed = {str(period or "").strip() for period in (history_periods or []) if str(period or "").strip()}
    metric_cols = [
        ("Revenue", header_map.get("revenue")),
        ("Operating income", header_map.get("op_income")),
        ("Net income", header_map.get("net_income")),
        ("CFO", header_map.get("cfo")),
        ("Capex", header_map.get("capex")),
        ("Cash", header_map.get("cash")),
        ("Debt", header_map.get("debt")),
        ("Diluted EPS", header_map.get("eps_diluted") or header_map.get("eps")),
    ]
    records: List[QuarterNarrativeRecord] = []
    ticker_txt = str(ticker or "").strip().upper()
    for rr in range(2, int(ws.max_row or 0) + 1):
        period = _quarter_narrative_period_from_label_or_date(ws.cell(rr, label_col).value, ticker=ticker_txt)
        if not period or (allowed and period not in allowed):
            continue
        parts: List[str] = []
        amount: Any = ""
        unit = ""
        for label, col in metric_cols:
            if not col:
                continue
            val = ws.cell(rr, col).value
            if val in (None, ""):
                continue
            display = _quarter_narrative_format_surface_value(val, label=label, raw_dollars=True)
            if not display:
                continue
            parts.append(f"{label} {display}")
            if amount == "":
                amount = val
                unit = "$m" if "$" in display else ""
            if len(parts) >= 4:
                break
        if not parts:
            continue
        row_like = {"topic": "Quarter actuals", "headline": "Quarterly actuals context", "body": " ".join(parts)}
        why, model, valuation, guardrail = _quarter_narrative_implications_for_row(
            ticker_txt,
            row_like,
            "Earnings / margin",
            "Quarterly actuals context",
        )
        records.append(
            QuarterNarrativeRecord(
                ticker=ticker_txt,
                fiscal_period=period,
                source_period=period,
                source_date=_quarter_narrative_source_date_from_period(period),
                source_type="History_Q",
                source_note="Workbook History_Q actuals",
                category="Earnings / margin",
                theme="Quarterly actuals context",
                what_happened=f"History_Q contains source-backed actuals: {', '.join(parts)}.",
                management_framing="Reported quarter data is available in the model history table.",
                why_it_matters=why,
                model_implication=model,
                valuation_implication=valuation,
                double_count_guardrail=guardrail,
                linked_sheet="History_Q; Valuation; Quarter_Notes_UI",
                linked_metric="Quarterly actuals",
                amount=amount,
                unit=unit,
                confidence="medium",
                include_in_quarter_notes=True,
                include_in_investment_case=True,
            )
        )
    return records


def _quarter_narrative_records_from_operating_drivers(
    wb: Workbook,
    ticker: Any,
    *,
    history_periods: Optional[Sequence[str]] = None,
    max_per_period: int = 3,
) -> List[QuarterNarrativeRecord]:
    if "Operating_Drivers" not in getattr(wb, "sheetnames", []):
        return []
    ws = wb["Operating_Drivers"]
    ticker_txt = str(ticker or "").strip().upper()
    allowed = {str(period or "").strip() for period in (history_periods or []) if str(period or "").strip()}
    terms = _quarter_narrative_surface_row_terms(ticker_txt)
    quarter_cols: Dict[int, str] = {}
    by_period_count: Dict[str, int] = {}
    records: List[QuarterNarrativeRecord] = []
    for rr in range(1, int(ws.max_row or 0) + 1):
        row_labels: Dict[int, str] = {}
        for cc in range(1, int(ws.max_column or 0) + 1):
            period = _quarter_narrative_period_from_label_or_date(ws.cell(rr, cc).value, ticker=ticker_txt)
            if period:
                row_labels[cc] = period
        if row_labels and str(ws.cell(rr, 1).value or "").strip().lower() in {"quarter", "period", "metric / segment"}:
            quarter_cols = row_labels
            continue
        if row_labels and not quarter_cols:
            quarter_cols = row_labels
            continue
        label = _quarter_narrative_clean_text(ws.cell(rr, 1).value, max_chars=90)
        if not label or label.lower() in {"quarter", "metric / segment", "watch item", "current watchlist"}:
            continue
        label_low = label.lower()
        if not any(term in label_low for term in terms):
            continue
        for cc, period in quarter_cols.items():
            if allowed and period not in allowed:
                continue
            if by_period_count.get(period, 0) >= max(1, int(max_per_period or 1)):
                continue
            raw_value = ws.cell(rr, cc).value
            if raw_value in (None, ""):
                continue
            value_txt = _quarter_narrative_format_surface_value(raw_value, label=label, raw_dollars=False)
            if not value_txt:
                continue
            row_like = {"topic": label, "headline": label, "body": value_txt}
            category, theme = _quarter_narrative_category_theme(row_like)
            why, model, valuation, guardrail = _quarter_narrative_implications_for_row(ticker_txt, row_like, category, theme)
            linked_sheet = "Operating_Drivers; Quarter_Notes_UI"
            linked_metric = label
            if ticker_txt == "PBI" and any(token in label_low for token in ("presort", "sendtech")):
                model = (
                    "Use Segment Scenario Inputs for Presort/SendTech revenue and margin sensitivity; "
                    "keep it separate from manual bridge rows."
                )
                guardrail = (
                    "Do not double-enter the same Presort/SendTech uplift through both "
                    "Segment Scenario Inputs and manual bridge rows."
                )
                linked_sheet = (
                    "Operating_Drivers; Investment_Case; Scenario_Driver_Assumptions; "
                    "Quarter_Notes_UI"
                )
                linked_metric = f"Segment Scenario Inputs - {label}"
            records.append(
                QuarterNarrativeRecord(
                    ticker=ticker_txt,
                    fiscal_period=period,
                    source_period=period,
                    source_date=_quarter_narrative_source_date_from_period(period),
                    source_type="Operating_Drivers",
                    source_note="Rendered Operating_Drivers row",
                    category=category,
                    theme=theme or label,
                    what_happened=f"Operating_Drivers shows {label} at {value_txt} for {period}.",
                    management_framing="Driver row is already source-backed or model-derived in Operating_Drivers.",
                    why_it_matters=why,
                    model_implication=model,
                    valuation_implication=valuation,
                    double_count_guardrail=guardrail,
                    linked_sheet=linked_sheet,
                    linked_metric=linked_metric,
                    amount=_quarter_narrative_amount_from_surface_value(raw_value),
                    unit="",
                    confidence="medium",
                    include_in_quarter_notes=True,
                    include_in_investment_case=any(token in label_low for token in ("45z", "margin", "ebit", "fcf", "debt", "cost", "sales", "revenue")),
                )
            )
            by_period_count[period] = by_period_count.get(period, 0) + 1
    return records


def _quarter_narrative_records_from_promise_progress_ui(
    wb: Workbook,
    ticker: Any,
    *,
    history_periods: Optional[Sequence[str]] = None,
    max_per_period: int = 3,
) -> List[QuarterNarrativeRecord]:
    if "Promise_Progress_UI" not in getattr(wb, "sheetnames", []):
        return []
    ws = wb["Promise_Progress_UI"]
    ticker_txt = str(ticker or "").strip().upper()
    allowed = {str(period or "").strip() for period in (history_periods or []) if str(period or "").strip()}
    active_header: Dict[str, int] = {}
    by_period_count: Dict[str, int] = {}
    records: List[QuarterNarrativeRecord] = []
    for rr in range(1, int(ws.max_row or 0) + 1):
        header_map: Dict[str, int] = {}
        for cc in range(1, min(int(ws.max_column or 0), 15) + 1):
            label = str(ws.cell(rr, cc).value or "").strip().lower()
            if label:
                header_map[label] = cc
        if "metric" in header_map and ("new/current guide" in header_map or "target / plan" in header_map):
            active_header = header_map
            continue
        if not active_header:
            continue
        metric_col = active_header.get("metric") or active_header.get("milestone")
        metric = _quarter_narrative_clean_text(ws.cell(rr, int(metric_col or 1)).value, max_chars=80)
        if not metric or metric.lower() in {"metric", "milestone"}:
            continue
        stated_col = active_header.get("stated in")
        stated_period = _quarter_narrative_period_from_label_or_date(ws.cell(rr, int(stated_col or 0)).value if stated_col else "", ticker=ticker_txt)
        if not stated_period:
            section_period = ""
            for back in range(rr - 1, 0, -1):
                section_period = _quarter_narrative_period_from_label_or_date(ws.cell(back, 1).value, ticker=ticker_txt)
                if section_period:
                    break
            stated_period = section_period
        if not stated_period or (allowed and stated_period not in allowed):
            continue
        if by_period_count.get(stated_period, 0) >= max(1, int(max_per_period or 1)):
            continue
        new_col = active_header.get("new/current guide") or active_header.get("target / plan") or active_header.get("current guide")
        actual_col = active_header.get("actual")
        progress_col = active_header.get("progress / run-rate")
        status_col = active_header.get("status")
        horizon_col = active_header.get("horizon")
        source_col = active_header.get("source / note") or active_header.get("notes/source")
        guide = _quarter_narrative_clean_text(ws.cell(rr, int(new_col or 0)).value if new_col else "", max_chars=100)
        actual = _quarter_narrative_clean_text(ws.cell(rr, int(actual_col or 0)).value if actual_col else "", max_chars=70)
        progress = _quarter_narrative_clean_text(ws.cell(rr, int(progress_col or 0)).value if progress_col else "", max_chars=90)
        status = _quarter_narrative_clean_text(ws.cell(rr, int(status_col or 0)).value if status_col else "", max_chars=40)
        horizon = _quarter_narrative_clean_text(ws.cell(rr, int(horizon_col or 0)).value if horizon_col else "", max_chars=40)
        source_note = _quarter_narrative_clean_text(ws.cell(rr, int(source_col or 0)).value if source_col else "", max_chars=110)
        if not any((guide, actual, progress, status)):
            continue
        row_like = {"topic": metric, "headline": metric, "body": " ".join(part for part in (guide, actual, progress, status) if part)}
        why, model, valuation, guardrail = _quarter_narrative_implications_for_row(
            ticker_txt,
            row_like,
            "Guidance / promise",
            metric,
        )
        pieces = []
        if guide:
            pieces.append(f"guide {guide}")
        if actual:
            pieces.append(f"actual {actual}")
        if progress:
            pieces.append(f"progress {progress}")
        if status:
            pieces.append(f"status {status}")
        records.append(
            QuarterNarrativeRecord(
                ticker=ticker_txt,
                fiscal_period=stated_period,
                source_period=stated_period,
                source_date=_quarter_narrative_source_date_from_period(stated_period),
                source_type="Promise_Progress_UI",
                source_note=source_note,
                category="Guidance / promise",
                theme=metric,
                what_happened=f"{metric}: {'; '.join(pieces)}.",
                management_framing=f"Horizon {horizon}." if horizon else "",
                why_it_matters=why,
                model_implication=model,
                valuation_implication=valuation,
                double_count_guardrail=guardrail,
                linked_sheet="Promise_Progress_UI; Quarter_Notes_UI",
                linked_metric=metric,
                confidence="medium",
                include_in_quarter_notes=True,
                include_in_promise_progress=True,
                include_in_investment_case=True,
            )
        )
        by_period_count[stated_period] = by_period_count.get(stated_period, 0) + 1
    return records


def _quarter_narrative_records_from_workbook_surfaces(
    wb: Workbook,
    ticker: Any,
    *,
    history_periods: Optional[Sequence[str]] = None,
    max_per_period: int = 7,
) -> List[QuarterNarrativeRecord]:
    """Create narrative records from clean workbook surfaces already rendered.

    This bridges the gap where source-backed facts made it into History_Q,
    Operating_Drivers or Promise_Progress_UI but did not have a hand-curated
    narrative record.  It intentionally avoids raw transcript text and only
    promotes concise, already-modeled workbook facts.
    """
    records: List[QuarterNarrativeRecord] = []
    records.extend(_quarter_narrative_records_from_history_q(wb, ticker, history_periods=history_periods))
    records.extend(
        _quarter_narrative_records_from_operating_drivers(
            wb,
            ticker,
            history_periods=history_periods,
            max_per_period=3,
        )
    )
    records.extend(
        _quarter_narrative_records_from_promise_progress_ui(
            wb,
            ticker,
            history_periods=history_periods,
            max_per_period=3,
        )
    )
    out: List[QuarterNarrativeRecord] = []
    seen: Set[Tuple[str, str, str, str]] = set()
    by_period_count: Dict[str, int] = {}
    for rec in records:
        period = str(rec.fiscal_period or "").strip()
        if not period:
            continue
        if by_period_count.get(period, 0) >= max(1, int(max_per_period or 1)):
            continue
        key = (
            period.lower(),
            str(rec.source_type or "").strip().lower(),
            str(rec.theme or rec.linked_metric or "").strip().lower(),
            str(rec.what_happened or "").strip().lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(rec)
        by_period_count[period] = by_period_count.get(period, 0) + 1
    return out

def _quarter_narrative_records_for_context(
    ticker: Any,
    *,
    workbook: Optional[Workbook] = None,
    quarter_notes: Any = None,
    history_periods: Optional[Sequence[str]] = None,
    max_per_period: int = 5,
) -> List[QuarterNarrativeRecord]:
    ticker_txt = str(ticker or "").strip().upper()
    base_records = list(_quarter_narrative_records_for_ticker(ticker_txt))
    generated_records = _quarter_narrative_records_from_quarter_notes(
        ticker_txt,
        quarter_notes,
        history_periods=history_periods,
        max_per_period=max_per_period,
    )
    surface_records = (
        _quarter_narrative_records_from_workbook_surfaces(
            workbook,
            ticker_txt,
            history_periods=history_periods,
            max_per_period=max(6, int(max_per_period or 5) + 2),
        )
        if workbook is not None
        else []
    )
    def _theme_identity(rec: QuarterNarrativeRecord) -> str:
        blob = " ".join(
            str(part or "")
            for part in (rec.theme, rec.what_happened, rec.linked_metric, rec.model_implication)
        ).lower()
        if ticker_txt == "PBI":
            if "sendtech" in blob:
                return "sendtech solutions"
            if "presort" in blob:
                return "presort services"
        return re.sub(r"\s+", " ", str(rec.theme or rec.linked_metric or "").strip().lower())

    def _record_quality(rec: QuarterNarrativeRecord) -> Tuple[int, int]:
        confidence_rank = {"high": 3, "medium": 2, "med": 2, "low": 1}
        source_rank = {
            "manual": 5,
            "curated": 5,
            "earnings release": 4,
            "guidance profile": 4,
            "quarter_notes": 4,
            "promise": 3,
            "operating_drivers": 3,
            "history_q": 2,
        }
        source_blob = str(rec.source_type or "").strip().lower()
        source_score = max((rank for token, rank in source_rank.items() if token in source_blob), default=1)
        text_score = sum(
            1
            for value in (rec.what_happened, rec.why_it_matters, rec.model_implication, rec.double_count_guardrail)
            if str(value or "").strip()
        )
        quality = (
            source_score
            + confidence_rank.get(str(rec.confidence or "").strip().lower(), 0)
            + text_score
            + int(bool(str(rec.source_date or "").strip()))
            + int(bool(rec.include_in_investment_case))
            + int(bool(rec.include_in_promise_progress))
        )
        text_len = len(" ".join(str(value or "") for value in (rec.what_happened, rec.why_it_matters, rec.model_implication)))
        return quality, text_len

    winners: Dict[Tuple[str, str], QuarterNarrativeRecord] = {}
    exact_seen: Set[Tuple[str, str, str]] = set()
    for rec in [*base_records, *generated_records, *surface_records]:
        exact_key = (
            str(rec.fiscal_period or "").strip().lower(),
            str(rec.theme or rec.linked_metric or "").strip().lower(),
            str(rec.what_happened or "").strip().lower(),
        )
        if exact_key in exact_seen:
            continue
        exact_seen.add(exact_key)
        key = (str(rec.fiscal_period or "").strip().lower(), _theme_identity(rec))
        if not key[1]:
            key = (key[0], str(rec.what_happened or "").strip().lower()[:120])
        prev = winners.get(key)
        if prev is None or _record_quality(rec) > _record_quality(prev):
            winners[key] = rec
    return sorted(winners.values(), key=lambda rec: (_quarter_narrative_period_sort_key(rec.fiscal_period), str(rec.theme or "")), reverse=True)


def _quarter_narrative_read_block(records: Sequence[QuarterNarrativeRecord]) -> List[Tuple[str, str]]:
    if not records:
        missing = "No source-backed narrative items generated for this quarter."
        return [
            ("Model read", missing),
            ("What changed", missing),
            ("Watch next", "Add source-backed quarter narrative when clean company materials are available."),
            ("Key caveat", "No fake narrative was generated for this sparse quarter."),
        ]
    themes = [str(r.theme or "").strip() for r in records if str(r.theme or "").strip()]
    top_themes = ", ".join(themes[:3])
    promise_records = [r for r in records if r.include_in_promise_progress or "guidance" in str(r.category or "").lower()]
    margin_records = [r for r in records if "margin" in str(r.category or "").lower() or "margin" in str(r.theme or "").lower()]
    policy_records = [r for r in records if "policy" in str(r.category or "").lower() or "45z" in str(r.theme or "").lower()]
    first = records[0] if records else QuarterNarrativeRecord(ticker="", fiscal_period="")
    changed = " ".join(_quarter_narrative_compact_sentence(r.what_happened, max_chars=95) for r in records[:2])
    if promise_records:
        watch = f"Watch source-backed progress against {', '.join(str(r.theme or r.linked_metric) for r in promise_records[:2])}."
    elif policy_records:
        watch = "Watch facility, policy and cash-conversion proof points."
    elif margin_records:
        watch = "Watch whether margin bridge items convert into operating income."
    else:
        watch = "Watch whether the quarter's drivers become repeatable and modelable."
    caveat = next((str(r.double_count_guardrail or "").strip() for r in records if str(r.double_count_guardrail or "").strip()), "")
    return [
        ("Model read", _quarter_narrative_compact_sentence(f"{top_themes} drive the quarter read. {first.model_implication}", max_chars=260)),
        ("What changed", _quarter_narrative_compact_sentence(changed, max_chars=260)),
        ("Watch next", _quarter_narrative_compact_sentence(watch, max_chars=220)),
        ("Key caveat", _quarter_narrative_compact_sentence(caveat or "Use source-backed metrics and avoid stacking overlapping drivers.", max_chars=240)),
    ]


def _write_quarter_notes_ui_narrative_sheet(
    wb: Workbook,
    ticker: Any,
    records: Optional[Sequence[QuarterNarrativeRecord]] = None,
    *,
    quarters_shown: int = 12,
    history_periods: Optional[Sequence[str]] = None,
) -> bool:
    records_list = list(records if records is not None else _quarter_narrative_records_for_ticker(ticker))
    records_list = [r for r in records_list if r.include_in_quarter_notes]
    recent_history_periods = [
        str(period or "").strip()
        for period in (history_periods if history_periods is not None else _quarter_narrative_recent_history_periods(wb, limit=8))
        if str(period or "").strip()
    ][:8]
    if not records_list and not recent_history_periods:
        return False
    if "Quarter_Notes_UI" in wb.sheetnames:
        del wb["Quarter_Notes_UI"]
    ws = wb.create_sheet("Quarter_Notes_UI")

    max_col = 15
    blue = PatternFill("solid", fgColor="5B9BD5")
    sub_blue = PatternFill("solid", fgColor="DDEBF7")
    header_fill = PatternFill("solid", fgColor="EAF3F8")
    zebra_light = PatternFill("solid", fgColor="F7FBFF")
    zebra_dark = PatternFill("solid", fgColor="EDF4FB")
    white_font = Font(bold=True, size=14, color="FFFFFF")
    sub_font = Font(bold=True, size=13, color="1F4E78")
    body_font = Font(size=14, color="1F1F1F")
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    widths = {
        "A": 20,
        "B": 22,
        "C": 26,
        "D": 26,
        "E": 26,
        "F": 26,
        "G": 26,
        "H": 26,
        "I": 26,
        "J": 24,
        "K": 24,
        "L": 24,
        "M": 22,
        "N": 22,
        "O": 46,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    def _style_cells(row: int, fill: PatternFill, font: Optional[Font] = None, *, height: Optional[float] = None) -> None:
        for cc in range(1, max_col + 1):
            cell = ws.cell(row=row, column=cc)
            cell.fill = copy(fill)
            cell.border = copy(border)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if font is not None:
                cell.font = copy(font)
            elif cell.font is None:
                cell.font = copy(body_font)
        if height is not None:
            ws.row_dimensions[row].height = height

    def _merge_row(row: int, start_col: int, end_col: int) -> None:
        if end_col > start_col:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

    def _section(row: int, title: str) -> int:
        _style_cells(row, sub_blue, sub_font, height=25.0)
        ws.cell(row=row, column=1, value=title)
        _merge_row(row, 1, max_col)
        return row + 1

    def _table_header(row: int, placements: Sequence[Tuple[Any, ...]]) -> int:
        _style_cells(row, header_fill, Font(bold=True, size=13, color="000000"), height=26.0)
        for placement in placements:
            if len(placement) == 2:
                col, label = placement
                end_col = col
            else:
                col, end_col, label = placement[:3]
            ws.cell(row=row, column=col, value=label)
            _merge_row(row, int(col), int(end_col))
        return row + 1

    def _merge_and_write(row: int, start_col: int, end_col: int, value: Any) -> None:
        ws.cell(row=row, column=start_col, value=value)
        _merge_row(row, start_col, end_col)

    def _spacer(row: int, *, height: float = 12.0) -> int:
        for cc in range(1, max_col + 1):
            cell = ws.cell(row=row, column=cc)
            cell.value = None
            cell.fill = PatternFill("solid", fgColor="FFFFFF")
            cell.border = Border()
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = height
        return row + 1

    def _write_no_info_row(row: int, message: str = "No information") -> int:
        _style_cells(row, zebra_light, body_font, height=30.0)
        ws.cell(row=row, column=1, value=message)
        _merge_row(row, 1, 2)
        ws.cell(row=row, column=3, value="No source-backed narrative rows for this subsection.")
        _merge_row(row, 3, max_col - 3)
        ws.cell(row=row, column=max_col - 2, value="No source-backed rows")
        _merge_row(row, max_col - 2, max_col)
        return row + 1

    def _write_read_block(row: int, quarter_records: Sequence[QuarterNarrativeRecord]) -> int:
        row = _section(row, "Quarter read")
        for idx, (field, text) in enumerate(_quarter_narrative_read_block(quarter_records)):
            fill = zebra_light if idx % 2 == 0 else zebra_dark
            _style_cells(row, fill, body_font, height=_quarter_narrative_row_height(field, text, base=30.0, max_height=72.0))
            ws.cell(row=row, column=1, value=field).font = Font(bold=True, size=14, color="1F4E78")
            ws.cell(row=row, column=2, value=text)
            _merge_row(row, 2, max_col)
            row += 1
        return row

    def _write_key_developments(row: int, quarter_records: Sequence[QuarterNarrativeRecord]) -> int:
        row = _section(row, "Key developments")
        row = _table_header(
            row,
            [
                (1, 2, "Theme"),
                (3, 5, "What happened"),
                (6, 7, "Why it matters"),
                (8, 12, "Model / valuation implication"),
                (13, 15, "Source / confidence"),
            ],
        )
        if not quarter_records:
            return _write_no_info_row(row)
        for idx, rec in enumerate(quarter_records[:6]):
            fill = zebra_light if idx % 2 == 0 else zebra_dark
            model_text = " ".join(
                part for part in [rec.model_implication, rec.valuation_implication] if str(part or "").strip()
            )
            _style_cells(
                row,
                fill,
                body_font,
                height=_quarter_narrative_row_height(rec.what_happened, rec.why_it_matters, model_text),
            )
            _merge_and_write(row, 1, 2, rec.theme)
            _merge_and_write(row, 3, 5, _quarter_narrative_compact_sentence(rec.what_happened, max_chars=230))
            _merge_and_write(row, 6, 7, _quarter_narrative_compact_sentence(rec.why_it_matters, max_chars=170))
            _merge_and_write(row, 8, 12, _quarter_narrative_compact_sentence(model_text, max_chars=285))
            _merge_and_write(row, 13, 15, _quarter_narrative_source_label(rec))
            row += 1
        return row

    def _write_promise_interpretation(row: int, quarter_records: Sequence[QuarterNarrativeRecord]) -> int:
        promise_rows = [
            r
            for r in quarter_records
            if r.include_in_promise_progress or "guidance" in str(r.category or "").lower() or "promise" in str(r.category or "").lower()
        ]
        row = _section(row, "Guidance / Promise interpretation")
        row = _table_header(
            row,
            [
                (1, 2, "Promise / guidance item"),
                (3, 5, "Read"),
                (6, 8, "Actual / progress interpretation"),
                (9, 12, "Status / caveat"),
                (13, 15, "Source"),
            ],
        )
        if not promise_rows:
            return _write_no_info_row(row)
        for idx, rec in enumerate(promise_rows[:5]):
            fill = zebra_light if idx % 2 == 0 else zebra_dark
            read = rec.management_framing or rec.why_it_matters
            progress = rec.what_happened
            caveat = rec.double_count_guardrail or rec.confidence
            _style_cells(row, fill, body_font, height=_quarter_narrative_row_height(read, progress, caveat))
            _merge_and_write(row, 1, 2, rec.linked_metric or rec.theme)
            _merge_and_write(row, 3, 5, _quarter_narrative_compact_sentence(read, max_chars=210))
            _merge_and_write(row, 6, 8, _quarter_narrative_compact_sentence(progress, max_chars=210))
            _merge_and_write(row, 9, 12, _quarter_narrative_compact_sentence(caveat, max_chars=260))
            _merge_and_write(row, 13, 15, _quarter_narrative_source_label(rec))
            row += 1
        return row

    def _write_model_mapping(row: int, quarter_records: Sequence[QuarterNarrativeRecord]) -> int:
        mapping_rows = [
            r
            for r in quarter_records
            if r.include_in_investment_case or str(r.double_count_guardrail or "").strip() or str(r.linked_sheet or "").strip()
        ]
        row = _section(row, "Model mapping / double-count guardrails")
        row = _table_header(
            row,
            [
                (1, 2, "Driver"),
                (3, 6, "Model treatment"),
                (7, 12, "Double-count guardrail"),
                (13, 15, "Linked sheet / metric"),
            ],
        )
        if not mapping_rows:
            return _write_no_info_row(row)
        for idx, rec in enumerate(mapping_rows[:6]):
            fill = zebra_light if idx % 2 == 0 else zebra_dark
            linked = " | ".join(part for part in [rec.linked_sheet, rec.linked_metric] if str(part or "").strip())
            _style_cells(row, fill, body_font, height=_quarter_narrative_row_height(rec.model_implication, rec.double_count_guardrail, linked))
            _merge_and_write(row, 1, 2, rec.linked_metric or rec.theme)
            _merge_and_write(row, 3, 6, _quarter_narrative_compact_sentence(rec.model_implication, max_chars=260))
            _merge_and_write(row, 7, 12, _quarter_narrative_compact_sentence(rec.double_count_guardrail, max_chars=320))
            _merge_and_write(row, 13, 15, _quarter_narrative_compact_sentence(linked, max_chars=190))
            row += 1
        return row

    grouped: Dict[str, List[QuarterNarrativeRecord]] = {}
    for rec in records_list:
        grouped.setdefault(str(rec.fiscal_period or "").strip(), []).append(rec)
    ordered_periods = sorted(
        set(grouped.keys()).union(recent_history_periods),
        key=_quarter_narrative_period_sort_key,
        reverse=True,
    )[: max(1, int(quarters_shown or 12))]

    row = 1
    for block_idx, period in enumerate(ordered_periods):
        if block_idx:
            row = _spacer(row, height=24.0)
        quarter_records = grouped.get(period, [])
        _style_cells(row, blue, white_font, height=24.0)
        ws.cell(row=row, column=1, value=f"{period} - Quarter Notes")
        _merge_row(row, 1, max_col)
        row += 1
        row = _write_read_block(row, quarter_records)
        row = _spacer(row, height=10.0)
        row = _write_key_developments(row, quarter_records)
        row = _spacer(row, height=10.0)
        row = _write_promise_interpretation(row, quarter_records)
        row = _spacer(row, height=10.0)
        row = _write_model_mapping(row, quarter_records)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    return True
