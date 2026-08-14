"""Sector/general Investment Case workbook renderers."""
from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from typing import Any, Dict, List, MutableMapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .workbook_modules import ResolvedTickerModuleRoute


@dataclass(frozen=True)
class SectorInvestmentCaseRenderDeps:
    runtime: MutableMapping[str, Any]


def write_sector_investment_case_data_sheet(
    deps: SectorInvestmentCaseRenderDeps,
    df: Any,
    *,
    ticker: str,
) -> None:
    runtime = deps.runtime
    wb = runtime["wb"]
    data = df
    ticker_txt = str(ticker or "").strip().upper()
    if not ticker_txt:
        return
    sheet_name = f"{ticker_txt}_Investment_Case_Data"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    df = data.copy() if isinstance(data, pd.DataFrame) else pd.DataFrame()
    if df.empty:
        ws["A1"] = f"No {ticker_txt} investment-case data available."
        return
    cols = [
        "section",
        "metric",
        "value",
        "unit",
        "display",
        "source",
        "source_note",
        "investment_read",
        "current_trend",
        "beat_miss_risk",
        "bull_evidence",
        "bear_evidence",
        "next_proof_point",
        "current_read",
        "earnings_metric",
        "multiple_yield",
        "implied_value_share",
        "scenario_read",
        "cash_flag",
        "recurring_flag",
        "quality_read",
    ]
    cols = [c for c in cols if c in df.columns] + [c for c in df.columns if c not in cols]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Border(left=Side(style="thin", color="D9E2EA"), right=Side(style="thin", color="D9E2EA"), top=Side(style="thin", color="D9E2EA"), bottom=Side(style="thin", color="D9E2EA"))
    for cc, col in enumerate(cols, 1):
        cell = ws.cell(1, cc, col)
        cell.font = Font(bold=True, color="1F2933")
        cell.fill = header_fill
        cell.border = thin
    for rr, rec in enumerate(df[cols].to_dict("records"), 2):
        for cc, col in enumerate(cols, 1):
            value = rec.get(col)
            if isinstance(value, str):
                value = ILLEGAL_CHARACTERS_RE.sub("", value)
            cell = ws.cell(rr, cc, value)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=col in {"display", "source_note", "investment_read", "bull_evidence", "bear_evidence", "next_proof_point", "current_read", "scenario_read", "quality_read"})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    for cc, col in enumerate(cols, 1):
        width = 18
        if col in {"section", "metric"}:
            width = 30
        elif col in {"display", "source_note", "investment_read", "bull_evidence", "bear_evidence", "next_proof_point", "current_read", "scenario_read", "quality_read"}:
            width = 46
        ws.column_dimensions[get_column_letter(cc)].width = width


def write_sector_investment_case_sheet(
    deps: SectorInvestmentCaseRenderDeps,
    df: Any,
    *,
    ticker: str,
    profile_route: ResolvedTickerModuleRoute,
) -> None:
    runtime = deps.runtime
    wb = runtime["wb"]
    data = df
    SCENARIO_DRIVER_CAPITAL_STRUCTURE_INTEREST = runtime["SCENARIO_DRIVER_CAPITAL_STRUCTURE_INTEREST"]
    SCENARIO_DRIVER_CASH_FLOW_CAPEX = runtime["SCENARIO_DRIVER_CASH_FLOW_CAPEX"]
    SCENARIO_DRIVER_MANUAL_INCREMENTAL = runtime["SCENARIO_DRIVER_MANUAL_INCREMENTAL"]
    SCENARIO_DRIVER_MARGIN_EBITDA = runtime["SCENARIO_DRIVER_MARGIN_EBITDA"]
    SCENARIO_DRIVER_TAX_CREDIT_SUBSIDY = runtime["SCENARIO_DRIVER_TAX_CREDIT_SUBSIDY"]
    SCENARIO_TAX_CASH_ONLY = runtime["SCENARIO_TAX_CASH_ONLY"]
    SCENARIO_TAX_NON_TAXABLE_CREDIT = runtime["SCENARIO_TAX_NON_TAXABLE_CREDIT"]
    SCENARIO_TAX_NO_EPS_IMPACT = runtime["SCENARIO_TAX_NO_EPS_IMPACT"]
    SCENARIO_TAX_TAXABLE = runtime["SCENARIO_TAX_TAXABLE"]
    _ScenarioDriverBridgeSpec = runtime["_ScenarioDriverBridgeSpec"]
    _SegmentScenarioInputSpec = runtime["_SegmentScenarioInputSpec"]
    _bs_segments_latest_segment_margin_from_workbook = runtime["_bs_segments_latest_segment_margin_from_workbook"]
    _company_operating_margin_proxy_from_workbook = runtime["_company_operating_margin_proxy_from_workbook"]
    _date_or_none = runtime["_date_or_none"]
    _excel_manual_percent_active_formula = runtime["_excel_manual_percent_active_formula"]
    _excel_percent_value_expr = runtime["_excel_percent_value_expr"]
    _excel_visible_value_range_formula = runtime["_excel_visible_value_range_formula"]
    _fiscal_profile_from_workbook = runtime["_fiscal_profile_from_workbook"]
    _guidance_source_contract_label = runtime["_guidance_source_contract_label"]
    _history_q_latest_full_year_actuals_from_workbook = runtime["_history_q_latest_full_year_actuals_from_workbook"]
    _history_q_latest_full_year_period_set = runtime["_history_q_latest_full_year_period_set"]
    _history_q_year_default_formulas = runtime["_history_q_year_default_formulas"]
    _operating_driver_latest_full_year_sum_from_workbook = runtime["_operating_driver_latest_full_year_sum_from_workbook"]
    _operating_driver_ttm_sum_from_workbook = runtime["_operating_driver_ttm_sum_from_workbook"]
    _scenario_bridge_eps_value_formula = runtime["_scenario_bridge_eps_value_formula"]
    _scenario_bridge_row_values = runtime["_scenario_bridge_row_values"]
    _segment_scenario_specs_from_records = runtime["_segment_scenario_specs_from_records"]
    _shared_visible_period_text = runtime["_shared_visible_period_text"]
    _write_scenario_bridge_tax_treatment_sheet = runtime["_write_scenario_bridge_tax_treatment_sheet"]
    _write_scenario_driver_assumptions_sheet = runtime["_write_scenario_driver_assumptions_sheet"]
    ticker_txt = str(ticker or "").strip().upper()
    if not ticker_txt:
        return
    if profile_route.ticker != ticker_txt:
        raise ValueError(
            f"Investment-case profile route ticker {profile_route.ticker!r} does not match {ticker_txt!r}."
        )
    profile_pack_ids = frozenset(profile_route.profile_pack_ids)
    has_pbi_profile_packs = {"shipping_mail_pack", "bank_pack"} <= profile_pack_ids
    has_commodity_ethanol_pack = "commodity_ethanol_pack" in profile_pack_ids
    sheet_name = f"{ticker_txt}_Investment_Case"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    df = data.copy() if isinstance(data, pd.DataFrame) else pd.DataFrame()
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    ws.sheet_view.zoomScale = 112
    title_fill = PatternFill("solid", fgColor="4472C4")
    section_fill = PatternFill("solid", fgColor="5B9BD5")
    header_fill = PatternFill("solid", fgColor="EAF3F8")
    subheader_fill = PatternFill("solid", fgColor="DDEBF7")
    alt_fill = PatternFill("solid", fgColor="F7FAFC")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    callout_fill = PatternFill("solid", fgColor="F2F6FA")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Border(left=Side(style="thin", color="D9E2EA"), right=Side(style="thin", color="D9E2EA"), top=Side(style="thin", color="D9E2EA"), bottom=Side(style="thin", color="D9E2EA"))
    dark = "1F2933"
    max_col = 10

    def _records(section: str) -> List[Dict[str, Any]]:
        if df.empty or "section" not in df.columns:
            return []
        return df[df["section"].astype(str).eq(section)].to_dict("records")

    def _section(row: int, title: str) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        cell = ws.cell(row, 1, title)
        cell.fill = section_fill
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for cc in range(1, max_col + 1):
            ws.cell(row, cc).fill = section_fill
            ws.cell(row, cc).border = thin
        ws.row_dimensions[row].height = 21
        return row + 1

    def _header(row: int, labels: Sequence[str], merge_spans: Sequence[Tuple[int, int]] = ()) -> int:
        for cc in range(1, max_col + 1):
            val = labels[cc - 1] if cc <= len(labels) else ""
            cell = ws.cell(row, cc, val)
            cell.fill = header_fill
            cell.font = Font(bold=True, size=12, color=dark)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center")
        _merge(row, merge_spans)
        ws.row_dimensions[row].height = 21
        return row + 1

    def _merge(row: int, spans: Sequence[Tuple[int, int]]) -> None:
        for first, last in spans:
            if last > first:
                try:
                    ws.merge_cells(start_row=row, start_column=first, end_row=row, end_column=last)
                except Exception:
                    pass
                ws.cell(row, first).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _row(row: int, vals: Sequence[Any], *, fill: Optional[PatternFill] = None, spans: Sequence[Tuple[int, int]] = ()) -> int:
        row_fill = fill or white_fill
        for cc in range(1, max_col + 1):
            val = vals[cc - 1] if cc <= len(vals) else ""
            if isinstance(val, str):
                val = _shared_visible_period_text(ILLEGAL_CHARACTERS_RE.sub("", val))
            cell = ws.cell(row, cc, val)
            cell.fill = copy(row_fill)
            cell.border = thin
            cell.font = Font(size=12, color=dark)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc in {1, 2, 3, 4, 5})
        _merge(row, spans)
        ws.row_dimensions[row].height = 21
        return row + 1

    def _extract_display_part(display: Any, pattern: str) -> str:
        txt = str(display or "")
        m = re.search(pattern, txt, flags=re.I)
        return m.group(1).strip() if m else ""

    def _source_read(rec: Dict[str, Any]) -> str:
        return str(rec.get("investment_read") or rec.get("source_note") or rec.get("source") or "").strip()

    def _num(value: Any) -> Optional[float]:
        if value is None or pd.isna(value):
            return None
        try:
            return float(value)
        except Exception:
            return None

    def _numeric_value(section: str, metric: str) -> Optional[float]:
        for rec in _records(section):
            if str(rec.get("metric") or "").strip().lower() == str(metric or "").strip().lower():
                return _num(rec.get("value"))
        return None

    def _manual_period_labels() -> Tuple[str, str, str]:
        return "2025 year", "2026-Q2", "2026 year"

    def _manual_input_specs() -> List[Tuple[str, str, Any, Any, Any, Any, str, str]]:
        history_period_set = _history_q_latest_full_year_period_set(wb, ticker=ticker_txt)
        if not history_period_set.get("quarter_criteria"):
            full_year_label = _manual_period_labels()[0]
            year_match = re.search(r"\b(20\d{2})\b", str(full_year_label or ""))
            profile = _fiscal_profile_from_workbook(wb, ticker=ticker_txt)
            if year_match and profile.year_end_month == 12 and profile.year_end_day == 31 and profile.year_label == "end":
                fy = int(year_match.group(1))
                history_period_set = dict(history_period_set)
                history_period_set["fiscal_year"] = fy
                history_period_set["quarter_criteria"] = [
                    date(fy, 3, 31),
                    date(fy, 6, 30),
                    date(fy, 9, 30),
                    date(fy, 12, 31),
                ]
                history_period_set["previous_quarter_criteria"] = [
                    date(fy - 1, 3, 31),
                    date(fy - 1, 6, 30),
                    date(fy - 1, 9, 30),
                    date(fy - 1, 12, 31),
                ]
                history_period_set["quarter_dates"] = list(history_period_set["quarter_criteria"])
                history_period_set["previous_quarter_dates"] = list(history_period_set["previous_quarter_criteria"])
        history_fy = _history_q_latest_full_year_actuals_from_workbook(wb, ticker=ticker_txt)
        history_fy_formulas = _history_q_year_default_formulas(
            (2025, 1, 1),
            (2026, 1, 1),
            fiscal_year=history_period_set.get("fiscal_year"),
            quarter_dates=history_period_set.get("quarter_dates"),
            previous_quarter_dates=history_period_set.get("previous_quarter_dates"),
            quarter_criteria=history_period_set.get("quarter_criteria"),
            previous_quarter_criteria=history_period_set.get("previous_quarter_criteria"),
        )

        def _fy_default(key: str) -> Any:
            val = history_fy.get(key)
            if val is None:
                return history_fy_formulas.get(key, '=""')
            if key in {"operating_margin", "tax_rate", "revenue_growth"}:
                return float(val)
            return round(float(val), 2)

        def _valuation_latest_row_value_formula(label: str) -> str:
            safe_label = str(label).replace('"', '""')
            row_expr = f'INDEX(Valuation!$B:$M,MATCH("{safe_label}",Valuation!$A:$A,0),0)'
            return f'=IFERROR(LOOKUP(2,1/({row_expr}<>""),{row_expr}),"")'

        def _valuation_latest_row_value_formula(label: str) -> str:
            if "Valuation" not in wb.sheetnames:
                return '=""'
            safe_label = str(label).replace('"', '""')
            row_expr = f'INDEX(Valuation!$B:$M,MATCH("{safe_label}",Valuation!$A:$A,0),0)'
            return f'=IFERROR(LOOKUP(2,1/({row_expr}<>""),{row_expr}),"")'

        def _adjusted_metrics_year_sum_formula(metric: str) -> str:
            safe_metric = str(metric).replace('"', '""')
            metric_range = f'INDEX(Adjusted_Metrics!$A:$Z,0,MATCH("{safe_metric}",Adjusted_Metrics!$1:$1,0))'
            date_range = f'INDEX(Adjusted_Metrics!$A:$Z,0,MATCH("quarter",Adjusted_Metrics!$1:$1,0))'
            criteria = list(history_period_set.get("quarter_criteria") or history_period_set.get("quarter_dates") or [])

            def _criteria_expr(value: Any) -> str:
                if isinstance(value, str):
                    return f'"{value.replace(chr(34), chr(34) + chr(34))}"'
                qd = _date_or_none(value)
                if qd is not None:
                    return f"DATE({qd.year},{qd.month},{qd.day})"
                return f'"{str(value or "").replace(chr(34), chr(34) + chr(34))}"'

            terms = [
                f"SUMIFS({metric_range},{date_range},{_criteria_expr(crit)})"
                for crit in criteria
                if crit not in (None, "")
            ]
            if terms:
                return f'=IFERROR(({"+".join(terms)})/1000000,"")'
            return (
                f'=IFERROR(SUMIFS({metric_range},{date_range},">="&DATE(2025,1,1),'
                f'{date_range},"<"&DATE(2026,1,1))/1000000,"")'
            )

        gpre_45z_fy = _operating_driver_latest_full_year_sum_from_workbook(wb, "45Z value realized ($m)") if has_commodity_ethanol_pack else None
        gpre_45z_ttm = _operating_driver_ttm_sum_from_workbook(wb, "45Z value realized ($m)") if has_commodity_ethanol_pack else None
        pbi_revenue_guidance_midpoint = 1830.0 if has_pbi_profile_packs else None
        pbi_eps_guidance_midpoint = 1.575 if has_pbi_profile_packs else None
        pbi_fcf_guidance_midpoint = 362.5 if has_pbi_profile_packs else None
        pbi_adjusted_ebit_guidance_midpoint = 445.0 if has_pbi_profile_packs else None
        pbi_cost_savings_run_rate = 157.0 if has_pbi_profile_packs else None
        pbi_cost_savings_target_midpoint = 190.0 if has_pbi_profile_packs else None
        scenario_tax_rate = _numeric_value("Assumptions", "Tax rate")
        if scenario_tax_rate is None:
            scenario_tax_rate = history_fy.get("tax_rate")
        if scenario_tax_rate is not None and float(scenario_tax_rate) > 1.0:
            scenario_tax_rate = float(scenario_tax_rate) / 100.0
        if scenario_tax_rate is None or not (0.0 <= float(scenario_tax_rate) <= 0.35):
            scenario_tax_rate = None
        scenario_tax_note = (
            "Clean model tax rate; 25% default scenario tax rate if unavailable."
            if scenario_tax_rate is not None
            else "Uses 25% default scenario tax rate if no clean source."
        )
        specs: List[Tuple[str, str, Any, Any, Any, Any, str, str]] = [
            ("price", "Current share price", '=""', '=""', '=""', '=""', "Manual input only.", "$0.00"),
            ("shares", "Diluted shares", '=IFERROR(IF(SharesDiluted<>"",SharesDiluted,Shares),"")', '=IFERROR(IF(SharesDiluted<>"",SharesDiluted,Shares),"")', '=""', '=""', "Model share denominator.", "#,##0.0"),
            ("net_debt", "Net debt / net cash", '=IFERROR(NetDebt,"")', '=IFERROR(NetDebt,"")', '=""', '=""', "Positive = net debt; negative = net cash.", "#,##0.0"),
            ("revenue", "Forward revenue", _fy_default("revenue_m"), '=IFERROR(Revenue_TTM,"")', pbi_revenue_guidance_midpoint if pbi_revenue_guidance_midpoint is not None else '=""', '=""', "TTM default; clean guide visible when available.", "#,##0.0"),
            ("eps", "Forward EPS", _fy_default("eps"), '=IFERROR(IF(Adj_EPS_TTM<>"",Adj_EPS_TTM,EPS_TTM),"")', pbi_eps_guidance_midpoint if pbi_eps_guidance_midpoint is not None else '=""', '=""', "Adjusted EPS preferred; guide visible when available.", "$0.00"),
            ("ebitda", "Forward Adj EBITDA", _fy_default("ebitda_m"), '=IFERROR(IF(ThesisBaseAdjEBITDA_FY<>"",ThesisBaseAdjEBITDA_FY,Adj_EBITDA),"")', '=""', '=""', "Adjusted EBITDA base.", "#,##0.0"),
            ("fcf", "Forward FCF", _fy_default("fcf_m"), '=IFERROR(FCF_TTM,"")' if has_pbi_profile_packs else '=IFERROR(IF(Adj_FCF_TTM<>"",Adj_FCF_TTM,FCF_TTM),"")', pbi_fcf_guidance_midpoint if pbi_fcf_guidance_midpoint is not None else '=""', '=""', "FCF base; guide visible when available.", "#,##0.0"),
            ("operating_margin", "Operating margin", _fy_default("operating_margin"), '=IFERROR(CompanyOperatingMargin_TTM,"")', '=""', '=""', "Operating income / revenue; proxy if needed.", "0.0%"),
            ("capex", "Capex", _fy_default("capex_m"), '=IFERROR(Capex_TTM,"")', "=20" if has_commodity_ethanol_pack else '=""', '=""', "Capex changes affect FCF only.", "#,##0.0"),
            ("pe", "P/E multiple", '=IFERROR(IF(Target_PE<>"",Target_PE,10),10)', '=IFERROR(IF(Target_PE<>"",Target_PE,10),10)', '=""', '=""', "Scenario P/E lens.", "0.0x"),
            ("ev_multiple", "EV/Adj EBITDA multiple", '=IFERROR(IF(Target_EV_AdjEBITDA<>"",Target_EV_AdjEBITDA,8),8)', '=IFERROR(IF(Target_EV_AdjEBITDA<>"",Target_EV_AdjEBITDA,8),8)', '=""', '=""', "Scenario EV/EBITDA lens.", "0.0x"),
            ("fcf_yield", "FCF yield", '=IFERROR(IF(Target_EV_Yield>1,Target_EV_Yield/100,Target_EV_Yield),0.07)', '=IFERROR(IF(Target_EV_Yield>1,Target_EV_Yield/100,Target_EV_Yield),0.07)', '=""', '=""', "Scenario FCF yield.", "0.0%"),
        ]
        if has_pbi_profile_packs:
            specs.extend(
                [
                    (
                        "adjusted_ebit",
                        "Adjusted EBIT",
                        _adjusted_metrics_year_sum_formula("adj_ebit"),
                        _valuation_latest_row_value_formula("Adj EBIT (TTM)"),
                        pbi_adjusted_ebit_guidance_midpoint if pbi_adjusted_ebit_guidance_midpoint is not None else '=""',
                        '=""',
                        "TTM default from Valuation; 2026 guide midpoint.",
                        "#,##0.0",
                    ),
                    ("debt_paydown", "Debt paydown", '=""', '=""', '=""', '=""', "Manual deleveraging input.", "#,##0.0"),
                    ("interest_refi", "Interest/refinancing cost", '=""', '=IFERROR(InterestPaid_TTM,"")', '=""', '=""', "Current interest burden default.", "#,##0.0"),
                    (
                        "cost_savings",
                        "Cost savings target / run-rate ($m)",
                        pbi_cost_savings_run_rate if pbi_cost_savings_run_rate is not None else '=""',
                        pbi_cost_savings_run_rate if pbi_cost_savings_run_rate is not None else '=""',
                        pbi_cost_savings_target_midpoint if pbi_cost_savings_target_midpoint is not None else '=""',
                        '=""',
                        "Run-rate baseline; target midpoint when clean.",
                        "#,##0.0",
                    ),
                ]
            )
        elif has_commodity_ethanol_pack:
            specs.extend(
                [
                    ("crush_margin", "Crush margin uplift ($m)", '=""', '=""', '=""', '=""', "Manual $m EBITDA uplift.", "#,##0.0"),
                    (
                        "credit_45z",
                        "45Z contribution / guide ($m)",
                        round(float(gpre_45z_fy), 1) if gpre_45z_fy is not None else '=""',
                        round(float(gpre_45z_ttm), 1) if gpre_45z_ttm is not None else '=""',
                        212.5,
                        '=""',
                        "Reported baseline plus clean guide midpoint.",
                        "#,##0.0",
                    ),
                    ("fcf_conversion", "FCF conversion", '=""', '=""', '=""', '=""', "Optional conversion input.", "0.0%"),
                    ("policy_upside", "Policy / RVO / E15 / export", '=""', '=""', '=""', '=""', "Manual $m uplift/drag.", "#,##0.0"),
                ]
            )
        specs.append(("scenario_tax_rate", "Scenario tax rate", _fy_default("tax_rate"), scenario_tax_rate if scenario_tax_rate is not None else '=""', '=""', '=""', scenario_tax_note, "0.0%"))
        return specs

    def _write_manual_inputs(row: int) -> Tuple[int, Dict[str, str]]:
        refs: Dict[str, str] = {}
        latest_year_label, next_q_label, current_year_label = _manual_period_labels()
        row = _section(row, "Manual Market / Scenario Inputs")
        row = _header(
            row,
            [
                "Input",
                f"Model default ({latest_year_label})",
                "Model default (TTM)",
                f"Guidance ({current_year_label})",
                f"Guidance ({next_q_label})",
                "Manual override",
                "Active value",
                "Notes",
            ],
            merge_spans=[(8, max_col)],
        )
        for idx, (key, label, fy_default, ttm_default, current_year_guidance, next_q_guidance, note, number_format) in enumerate(_manual_input_specs()):
            current_row = row
            if key == "price":
                active_formula = f'=IF(F{current_row}<>"",F{current_row},"")'
            elif key == "scenario_tax_rate":
                active_formula = (
                    f'=IF(F{current_row}<>"",F{current_row},IF(C{current_row}<>"",C{current_row},'
                    f'IF(B{current_row}<>"",B{current_row},IF(D{current_row}<>"",D{current_row},'
                    f'IF(E{current_row}<>"",E{current_row},0.25)))))'
                )
            elif key == "fcf_yield":
                active_formula = _excel_manual_percent_active_formula(current_row)
            elif key in {"cost_savings", "credit_45z", "capex"}:
                active_formula = (
                    f'=IF(F{current_row}<>"",F{current_row},'
                    f'IF(D{current_row}<>"",D{current_row},IF(C{current_row}<>"",C{current_row},'
                    f'IF(B{current_row}<>"",B{current_row},E{current_row}))))'
                )
            else:
                active_formula = f'=IF(F{current_row}<>"",F{current_row},IF(C{current_row}<>"",C{current_row},IF(B{current_row}<>"",B{current_row},IF(D{current_row}<>"",D{current_row},E{current_row}))))'
            row = _row(
                row,
                [label, fy_default, ttm_default, current_year_guidance, next_q_guidance, "", active_formula, note],
                fill=alt_fill if idx % 2 else white_fill,
                spans=[(8, max_col)],
            )
            for cc in range(2, 8):
                ws.cell(current_row, cc).number_format = number_format
                ws.cell(current_row, cc).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(current_row, 6).fill = copy(input_fill)
            refs[key] = f"$G${current_row}"
            refs[f"{key}__latest_year"] = f"$B${current_row}"
            refs[f"{key}__ttm"] = f"$C${current_row}"
            refs[f"{key}__ttm_value"] = ttm_default
            refs[f"{key}__guidance_current"] = f"$D${current_row}"
            refs[f"{key}__guidance_next"] = f"$E${current_row}"
            refs[f"{key}__override"] = f"$F${current_row}"
        return row + 1, refs

    def _segment_scenario_specs() -> List[_SegmentScenarioInputSpec]:
        if not has_pbi_profile_packs:
            return []
        default_margin_proxy, default_margin_basis = _company_operating_margin_proxy_from_workbook(wb)
        specs = _segment_scenario_specs_from_records(
            _records("Segment Scenario Inputs"),
            default_margin_proxy=default_margin_proxy,
            default_margin_basis=default_margin_basis or "Company operating margin proxy",
        )
        for spec in specs:
            segment_margin, segment_margin_basis = _bs_segments_latest_segment_margin_from_workbook(wb, spec.label)
            if segment_margin is not None:
                spec.margin_conversion = segment_margin
                spec.margin_basis = segment_margin_basis
                spec.source_note = segment_margin_basis
                spec.feeds_bridge = spec.baseline_revenue_m is not None
        if specs:
            return specs
        return [
            _SegmentScenarioInputSpec("Presort", "Segment / business line", revenue_basis="", source_note="Missing segment revenue"),
            _SegmentScenarioInputSpec("SendTech", "Segment / business line", revenue_basis="", source_note="Missing segment revenue"),
        ]

    def _write_segment_scenario_inputs(row: int) -> Tuple[int, Dict[str, str]]:
        specs = _segment_scenario_specs()
        if not has_pbi_profile_packs:
            _write_scenario_driver_assumptions_sheet(
                wb,
                ticker=ticker_txt,
                enabled=False,
                disabled_note=(
                    "GPRE segment scenario disabled; use ethanol, 45Z, crush and policy drivers."
                    if has_commodity_ethanol_pack
                    else "Segment scenario disabled until ticker-specific segment inputs are configured."
                ),
            )
            return row, {}
        row = _section(row, "Segment Scenario Inputs")
        row = _header(
            row,
            [
                "Segment / category",
                "Type",
                "Baseline revenue",
                "Revenue % change",
                "Revenue impact",
                "Operating margin",
                "EBITDA impact",
                "Feeds bridge?",
                "Notes",
            ],
            merge_spans=[(9, max_col)],
        )
        impact_start: Optional[int] = None
        impact_end: Optional[int] = None
        for idx, spec in enumerate(specs):
            current_row = row
            note = spec.source_note or ("Missing segment margin" if spec.margin_conversion is None else "Informational only")
            has_margin = spec.margin_conversion is not None
            ebitda_impact: Any = f'=IFERROR(IF(OR(D{current_row}="",F{current_row}=""),0,E{current_row}*F{current_row}),0)' if has_margin else ""
            feeds_formula: Any = (
                f'=IF(AND(C{current_row}<>"",D{current_row}<>"",F{current_row}<>""),"Yes","No")'
                if spec.feeds_bridge and has_margin and spec.baseline_revenue_m is not None
                else "No"
            )
            row = _row(
                row,
                [
                    spec.label,
                    spec.category_type,
                    spec.baseline_revenue_m if spec.baseline_revenue_m is not None else "",
                    "",
                    f'=IFERROR(IF(D{current_row}="",0,C{current_row}*D{current_row}),0)',
                    spec.margin_conversion if spec.margin_conversion is not None else "",
                    ebitda_impact,
                    feeds_formula,
                    note,
                ],
                fill=alt_fill if idx % 2 else white_fill,
                spans=[(9, max_col)],
            )
            impact_start = current_row if impact_start is None else impact_start
            impact_end = current_row
            sheet_ref = f"'{ws.title}'"
            spec.revenue_change_ref = f"={sheet_ref}!$D${current_row}"
            spec.revenue_impact_ref = f"={sheet_ref}!$E${current_row}"
            spec.ebitda_impact_ref = f"={sheet_ref}!$G${current_row}"
            spec.feeds_bridge_ref = f"={sheet_ref}!$H${current_row}"
            ws.cell(current_row, 3).number_format = "#,##0.0"
            ws.cell(current_row, 4).number_format = "0.0%"
            ws.cell(current_row, 4).fill = copy(input_fill)
            ws.cell(current_row, 5).number_format = "#,##0.0"
            ws.cell(current_row, 6).number_format = "0.0%"
            ws.cell(current_row, 7).number_format = "#,##0.0"
            ws.cell(current_row, 8).alignment = Alignment(horizontal="left", vertical="center")
            for cc in range(1, max_col + 1):
                ws.cell(current_row, cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc in {1, 2, 9})
        _write_scenario_driver_assumptions_sheet(wb, ticker=ticker_txt, segment_specs=specs, enabled=True)
        selected_formula = (
            f'=SUMIF(H{impact_start}:H{impact_end},"Yes",G{impact_start}:G{impact_end})'
            if impact_start is not None and impact_end is not None
            else ""
        )
        return row + 1, {"selected_segment_impact": selected_formula} if selected_formula else {}

    def _manual_ref(refs: Dict[str, str], key: str) -> str:
        return refs.get(key, '""')

    def _manual_part_ref(refs: Dict[str, str], key: str, part: str) -> str:
        return refs.get(f"{key}__{part}", '""')

    def _active_value_formula(ref: str) -> str:
        return f'=IF({ref}="","",{ref})'

    def _write_market_pricing(row: int, refs: Dict[str, str]) -> int:
        price = _manual_ref(refs, "price")
        shares = _manual_ref(refs, "shares")
        net_debt = _manual_ref(refs, "net_debt")
        eps = _manual_ref(refs, "eps")
        ebitda = _manual_ref(refs, "ebitda")
        fcf = _manual_ref(refs, "fcf")
        row = _section(row, "What Market Is Pricing")
        row = _header(row, ["Metric", "Value / read", "", "", "Notes"], merge_spans=[(2, 4), (5, max_col)])
        market_rows = [
            (
                "Market price input",
                f'=IF({price}="","Manual current share price needed to calculate market-implied expectations.","Manual share price active; implied multiples use the active scenario inputs.")',
                "Leave current share price blank when no manual market price is intended.",
                "@",
            ),
            ("Implied market cap", f'=IF({price}="","",{price}*{shares})', "Current share price x active share count.", "#,##0.0"),
            ("Implied EV", f'=IF(OR({price}="",{shares}="",{net_debt}=""),"",{price}*{shares}+{net_debt})', "Market cap plus net debt / less net cash.", "#,##0.0"),
            ("Implied P/E", f'=IFERROR(IF(OR({price}="",{eps}="",{eps}=0),"",{price}/{eps}),"")', "Price divided by active forward EPS.", "0.0x"),
            ("Implied EV/Adj EBITDA", f'=IFERROR(IF(OR({price}="",{shares}="",{ebitda}="",{ebitda}=0),"",({price}*{shares}+{net_debt})/{ebitda}),"")', "Implied enterprise value divided by active Adj EBITDA.", "0.0x"),
            ("Implied FCF yield", f'=IFERROR(IF(OR({price}="",{shares}="",{fcf}=""),"",{fcf}/({price}*{shares})),"")', "Active FCF divided by implied market cap.", "0.0%"),
        ]
        for idx, (metric, value, note, number_format) in enumerate(market_rows):
            current_row = row
            row = _row(row, [metric, value, "", "", note], fill=alt_fill if idx % 2 else white_fill, spans=[(2, 4), (5, max_col)])
            ws.cell(current_row, 2).number_format = number_format
            ws.cell(current_row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return row

    def _write_scenario_driver_bridge(row: int, refs: Dict[str, str], segment_refs: Optional[Dict[str, str]] = None) -> Tuple[int, Dict[str, str]]:
        eps = _manual_ref(refs, "eps")
        ebitda = _manual_ref(refs, "ebitda")
        fcf = _manual_ref(refs, "fcf")
        shares = _manual_ref(refs, "shares")
        eps_override = _manual_part_ref(refs, "eps", "override")
        shares_ttm = _manual_part_ref(refs, "shares", "ttm")
        scenario_tax_rate = _manual_ref(refs, "scenario_tax_rate")
        segment_refs = segment_refs or {}
        selected_segment_impact = segment_refs.get("selected_segment_impact", "")
        refs_out: Dict[str, str] = {}
        row = _section(row, "Scenario Driver Bridge")
        row = _row(
            row,
            ["Bridge-adjusted values start from active inputs and add incremental effects below | Taxable EPS impacts use active scenario tax rate."],
            fill=white_fill,
            spans=[(1, max_col)],
        )
        row = _header(
            row,
            [
                "Bridge item",
                "Baseline included",
                "Active / guide",
                "Incremental effect",
                "EPS impact",
                "EBITDA impact",
                "FCF impact",
                "Read",
            ],
            merge_spans=[(8, max_col)],
        )
        incremental_start = row
        after_tax_factor = None
        tax_source_basis = "Manual Market / Scenario Inputs active Scenario tax rate; defaults to 25% if no clean source."

        if has_pbi_profile_packs:
            cost_savings = _manual_ref(refs, "cost_savings")
            cost_savings_baseline = _manual_part_ref(refs, "cost_savings", "ttm")
            interest_refi = _manual_ref(refs, "interest_refi")
            interest_baseline = _manual_part_ref(refs, "interest_refi", "ttm")
            debt_paydown = _manual_ref(refs, "debt_paydown")
            capex = _manual_ref(refs, "capex")
            capex_baseline = _manual_part_ref(refs, "capex", "ttm")
            bridge_specs = [
                _ScenarioDriverBridgeSpec(
                    "Incremental cost savings vs baseline",
                    SCENARIO_DRIVER_MARGIN_EBITDA,
                    f'=IF({cost_savings_baseline}="","Unknown",{cost_savings_baseline})',
                    _active_value_formula(cost_savings),
                    "Run-rate baseline vs active target.",
                    ebitda_impact="same",
                    fcf_impact="none",
                    eps_impact="auto",
                    tax_treatment=SCENARIO_TAX_TAXABLE,
                    tax_source_basis="Operating cost savings assumed taxable; baseline is the visible run-rate/TTM savings row.",
                ),
                _ScenarioDriverBridgeSpec(
                    "Interest/refinancing effect vs baseline",
                    SCENARIO_DRIVER_CAPITAL_STRUCTURE_INTEREST,
                    f'=IF({interest_baseline}="","Unknown",{interest_baseline})',
                    _active_value_formula(interest_refi),
                    "Lower interest vs baseline lifts FCF; EPS needs tax conversion.",
                    reverse_incremental=True,
                    ebitda_impact="none",
                    fcf_impact="same",
                    eps_impact="auto",
                    tax_treatment=SCENARIO_TAX_TAXABLE,
                    tax_source_basis="Interest/refinancing is treated as pre-tax interest effect when tax conversion is available.",
                ),
                _ScenarioDriverBridgeSpec(
                    "Selected segment revenue/margin impact",
                    SCENARIO_DRIVER_MARGIN_EBITDA,
                    0,
                    selected_segment_impact or 0,
                    "Selected Segment Scenario Inputs feed taxable EBITDA uplift.",
                    explicit_incremental=True,
                    ebitda_impact="same",
                    fcf_impact="none",
                    eps_impact="auto",
                    tax_treatment=SCENARIO_TAX_TAXABLE,
                    tax_source_basis="Segment Scenario Inputs selected Feeds bridge? rows.",
                ),
                _ScenarioDriverBridgeSpec(
                    "Capex change vs baseline",
                    SCENARIO_DRIVER_CASH_FLOW_CAPEX,
                    f'=IF({capex_baseline}="","Unknown",{capex_baseline})',
                    _active_value_formula(capex),
                    "Capex change affects FCF only.",
                    ebitda_impact="none",
                    fcf_impact="negative",
                    eps_impact="none",
                    tax_treatment=SCENARIO_TAX_CASH_ONLY,
                    tax_source_basis="Capex affects cash flow, not direct EPS or Adj EBITDA.",
                ),
                _ScenarioDriverBridgeSpec(
                    "Debt paydown / net debt",
                    SCENARIO_DRIVER_CAPITAL_STRUCTURE_INTEREST,
                    _active_value_formula(_manual_ref(refs, "net_debt")),
                    _active_value_formula(debt_paydown),
                    "Use net debt override; no interest rate assumed.",
                    explicit_incremental=True,
                    ebitda_impact="none",
                    fcf_impact="none",
                    eps_impact="none",
                    tax_treatment=SCENARIO_TAX_NO_EPS_IMPACT,
                    tax_source_basis="Debt paydown changes net debt/equity bridge only unless a separate interest effect is modeled.",
                ),
            ]
        elif has_commodity_ethanol_pack:
            credit_45z = _manual_ref(refs, "credit_45z")
            crush_margin = _manual_ref(refs, "crush_margin")
            capex = _manual_ref(refs, "capex")
            capex_baseline = _manual_part_ref(refs, "capex", "ttm")
            policy_upside = _manual_ref(refs, "policy_upside")
            credit_45z_baseline = _manual_part_ref(refs, "credit_45z", "ttm")
            credit_45z_read = (
                "Incremental 45Z vs TTM Operating_Drivers baseline."
            )
            bridge_specs = [
                _ScenarioDriverBridgeSpec(
                    "Incremental 45Z uplift vs baseline",
                    SCENARIO_DRIVER_TAX_CREDIT_SUBSIDY,
                    f'=IF({credit_45z_baseline}="","Unknown",{credit_45z_baseline})',
                    _active_value_formula(credit_45z),
                    credit_45z_read,
                    ebitda_impact="same",
                    fcf_impact="none",
                    eps_impact="auto",
                    subsidy_basis="ebitda_like",
                    tax_treatment=SCENARIO_TAX_NON_TAXABLE_CREDIT,
                    tax_source_basis="45Z is source-backed as an EBITDA-like tax-credit contribution in Operating_Drivers/management guidance.",
                ),
                _ScenarioDriverBridgeSpec(
                    "Crush margin uplift ($m)",
                    SCENARIO_DRIVER_MARGIN_EBITDA,
                    0,
                    _active_value_formula(crush_margin),
                    "Direct manual EBITDA uplift.",
                    explicit_incremental=True,
                    ebitda_impact="same",
                    fcf_impact="none",
                    eps_impact="auto",
                    tax_treatment=SCENARIO_TAX_TAXABLE,
                    tax_source_basis="Crush margin uplift is an operating EBITDA uplift and taxable unless a tax-credit basis is documented.",
                ),
                _ScenarioDriverBridgeSpec(
                    "Capex change vs baseline",
                    SCENARIO_DRIVER_CASH_FLOW_CAPEX,
                    f'=IF({capex_baseline}="","Unknown",{capex_baseline})',
                    _active_value_formula(capex),
                    "Capex change affects FCF only.",
                    ebitda_impact="none",
                    fcf_impact="negative",
                    eps_impact="none",
                    tax_treatment=SCENARIO_TAX_CASH_ONLY,
                    tax_source_basis="Capex affects cash flow, not direct EPS or Adj EBITDA.",
                ),
                _ScenarioDriverBridgeSpec(
                    "Policy / RVO / E15 / export",
                    SCENARIO_DRIVER_MANUAL_INCREMENTAL,
                    0,
                    _active_value_formula(policy_upside),
                    "Explicit manual EBITDA uplift/drag.",
                    explicit_incremental=True,
                    ebitda_impact="same",
                    fcf_impact="none",
                    eps_impact="auto",
                    tax_treatment=SCENARIO_TAX_TAXABLE,
                    tax_source_basis="Numeric policy/RVO/E15/export input is treated as taxable operating uplift unless explicitly tax-like.",
                ),
            ]
        else:
            _write_scenario_bridge_tax_treatment_sheet(
                wb,
                ticker=ticker_txt,
                specs=(),
                after_tax_factor=after_tax_factor,
                tax_rate_ref=scenario_tax_rate,
                tax_source_basis=tax_source_basis,
            )
            row = _row(
                row,
                [
                    "No authorized profile-pack scenario drivers",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "Needs Review: ticker-specific scenario economics are unavailable until a declarative profile is registered.",
                ],
                fill=callout_fill,
                spans=[(8, max_col)],
            )
            return row + 1, {}

        _write_scenario_bridge_tax_treatment_sheet(
            wb,
            ticker=ticker_txt,
            specs=bridge_specs,
            after_tax_factor=after_tax_factor,
            tax_rate_ref=scenario_tax_rate,
            tax_source_basis=tax_source_basis,
        )
        incremental_rows = [
            _scenario_bridge_row_values(
                spec,
                row + idx,
                active_eps_ref=eps,
                active_ebitda_ref=ebitda,
                active_shares_ref=shares,
                baseline_shares_ref=shares_ttm,
                eps_override_ref=eps_override,
                after_tax_factor=after_tax_factor,
                tax_rate_ref=scenario_tax_rate,
            )
            for idx, spec in enumerate(bridge_specs)
        ]

        for idx, (item, baseline, active, incremental, eps_impact, ebitda_impact, fcf_impact, read) in enumerate(incremental_rows):
            current_row = row
            row = _row(
                row,
                [item, baseline, active, incremental, eps_impact, ebitda_impact, fcf_impact, read],
                fill=alt_fill if idx % 2 else white_fill,
                spans=[(8, max_col)],
            )
            for cc in range(2, 8):
                ws.cell(current_row, cc).number_format = "#,##0.0"
                ws.cell(current_row, cc).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(current_row, 5).number_format = "$0.00"
            ws.cell(current_row, 8).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        incremental_end = row - 1
        row += 1
        row = _header(row, ["Metric", "Active input", "", "Adjustment", "", "Bridge-adjusted value", "", "Read"], merge_spans=[(2, 3), (4, 5), (6, 7), (8, max_col)])
        summary_start = row
        bridge_eps_value = _scenario_bridge_eps_value_formula(
            summary_row=summary_start,
            eps_override_ref=eps_override,
            active_eps_ref=eps,
            active_shares_ref=shares,
            baseline_shares_ref=shares_ttm,
            eps_impact_start=incremental_start,
            eps_impact_end=incremental_end,
        )
        eps_adjustment = f'=IFERROR(IF(B{summary_start}="","",{bridge_eps_value[1:]}-B{summary_start}),"")'
        summary_rows = [
            (
                "Bridge EPS ($/sh)",
                _active_value_formula(eps),
                eps_adjustment,
                bridge_eps_value,
                "Share count and bridge impacts apply.",
                "bridge_eps",
                "$0.00",
            ),
            (
                "Bridge Adj EBITDA ($m)",
                _active_value_formula(ebitda),
                f'=IFERROR(SUM(F{incremental_start}:F{incremental_end}),"")',
                f'=IFERROR(IF(B{summary_start + 1}="","",B{summary_start + 1}+D{summary_start + 1}),"")',
                "Active EBITDA plus incremental EBITDA impacts.",
                "bridge_ebitda",
                "#,##0.0",
            ),
            (
                "Bridge FCF ($m)",
                _active_value_formula(fcf),
                f'=IFERROR(SUM(G{incremental_start}:G{incremental_end}),"")',
                f'=IFERROR(IF(B{summary_start + 2}="","",B{summary_start + 2}+D{summary_start + 2}),"")',
                "Active FCF plus incremental FCF impacts.",
                "bridge_fcf",
                "#,##0.0",
            ),
        ]

        for idx, (metric, active, adjustment, bridge_value, read, ref_key, number_format) in enumerate(summary_rows):
            current_row = row
            row = _row(
                row,
                [metric, active, "", adjustment, "", bridge_value, "", read],
                fill=alt_fill if idx % 2 else white_fill,
                spans=[(2, 3), (4, 5), (6, 7), (8, max_col)],
            )
            for cc in (2, 4, 6):
                ws.cell(current_row, cc).number_format = number_format
                ws.cell(current_row, cc).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(current_row, 8).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if ref_key:
                refs_out[ref_key] = f"$F${current_row}"
        return row + 1, refs_out

    def _write_manual_scenarios(row: int, refs: Dict[str, str], bridge_refs: Optional[Dict[str, str]] = None) -> int:
        bridge_refs = bridge_refs or {}
        eps = bridge_refs.get("bridge_eps") or _manual_ref(refs, "eps")
        ebitda = bridge_refs.get("bridge_ebitda") or _manual_ref(refs, "ebitda")
        fcf = bridge_refs.get("bridge_fcf") or _manual_ref(refs, "fcf")
        pe = _manual_ref(refs, "pe")
        ev_multiple = _manual_ref(refs, "ev_multiple")
        fcf_yield = _manual_ref(refs, "fcf_yield")
        fcf_yield_rate = _excel_percent_value_expr(fcf_yield)
        shares = _manual_ref(refs, "shares")
        net_debt = _manual_ref(refs, "net_debt")
        row = _section(row, "Bear / Base / Bull Scenario")
        row = _header(
            row,
            ["Scenario", "Key assumptions", "", "EPS", "Adj EBITDA", "FCF", "Value/share @ P/E", "Value/share @ EV/Adj EBITDA", "Value/share @ FCF yield", "Value range"],
            merge_spans=[(2, 3)],
        )
        scenarios = [
            ("Bear", "Lower earnings, lower multiple and higher FCF yield.", 0.90, 0.90, 0.85, 0.90, 0.90, 1.15),
            ("Base", "Active manual/default assumptions.", 1.00, 1.00, 1.00, 1.00, 1.00, 1.00),
            ("Bull", "Higher earnings, better multiple and lower FCF yield.", 1.10, 1.10, 1.15, 1.15, 1.15, 0.85),
        ]
        for idx, (name, assumptions, eps_factor, ebitda_factor, fcf_factor, pe_factor, ev_factor, yield_factor) in enumerate(scenarios):
            current_row = row
            eps_cell = f"D{current_row}"
            ebitda_cell = f"E{current_row}"
            fcf_cell = f"F{current_row}"
            row = _row(
                row,
                [
                    name,
                    assumptions,
                    "",
                    f'=IFERROR(IF({eps}="","",{eps}*{eps_factor}),"")',
                    f'=IFERROR(IF({ebitda}="","",{ebitda}*{ebitda_factor}),"")',
                    f'=IFERROR(IF({fcf}="","",{fcf}*{fcf_factor}),"")',
                    f'=IFERROR(IF(OR({eps_cell}="",{eps_cell}<=0,{pe}=""),"N/M",{eps_cell}*{pe}*{pe_factor}),"N/M")',
                    f'=IFERROR(IF(OR({ebitda_cell}="",{ev_multiple}="",{shares}="",{shares}=0),"",({ebitda_cell}*{ev_multiple}*{ev_factor}-{net_debt})/{shares}),"")',
                    f'=IFERROR(IF(OR({fcf_cell}="",{fcf_yield}="",{shares}="",{shares}=0),"",({fcf_cell}/(({fcf_yield_rate})*{yield_factor}))/{shares}),"")',
                    _excel_visible_value_range_formula(current_row),
                ],
                fill=alt_fill if idx % 2 else white_fill,
                spans=[(2, 3)],
            )
            for cc in (4, 7, 8, 9):
                ws.cell(current_row, cc).number_format = "$0.00"
            for cc in (5, 6):
                ws.cell(current_row, cc).number_format = "#,##0.0"
            ws.cell(current_row, 10).alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[current_row].height = 30
        row = _row(
            row,
            ["Uses Investment_Case manual inputs; may differ from Valuation Thesis Bridge."],
            fill=callout_fill,
            spans=[(1, max_col)],
        )
        return row

    def _write_manual_valuation_sensitivity(row: int, refs: Dict[str, str], bridge_refs: Optional[Dict[str, str]] = None) -> int:
        bridge_refs = bridge_refs or {}
        eps = bridge_refs.get("bridge_eps") or _manual_ref(refs, "eps")
        ebitda = bridge_refs.get("bridge_ebitda") or _manual_ref(refs, "ebitda")
        fcf = bridge_refs.get("bridge_fcf") or _manual_ref(refs, "fcf")
        pe = _manual_ref(refs, "pe")
        ev_multiple = _manual_ref(refs, "ev_multiple")
        fcf_yield = _manual_ref(refs, "fcf_yield")
        fcf_yield_rate = _excel_percent_value_expr(fcf_yield)
        shares = _manual_ref(refs, "shares")
        net_debt = _manual_ref(refs, "net_debt")
        row = _section(row, "Valuation Sensitivity")
        row = _header(row, ["EPS", "10x", "12x", "14x", "16x"])
        eps_factors = (0.90, 1.00, 1.10)
        for idx, factor in enumerate(eps_factors):
            current_row = row
            row = _row(
                row,
                [
                    f'=IFERROR(IF({eps}="","",{eps}*{factor}),"")',
                    f"=$A{current_row}*10",
                    f"=$A{current_row}*12",
                    f"=$A{current_row}*14",
                    f"=$A{current_row}*16",
                ],
                fill=alt_fill if idx % 2 else white_fill,
            )
            for cc in range(1, 6):
                ws.cell(current_row, cc).number_format = "$0.00"
                ws.cell(current_row, cc).alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        row = _header(row, ["Scenario", "EPS", "P/E", "Share price"])
        for idx, (name, eps_factor, pe_factor) in enumerate((("Bear", 0.90, 0.90), ("Base", 1.00, 1.00), ("Bull", 1.10, 1.15))):
            current_row = row
            row = _row(
                row,
                [
                    name,
                    f'=IFERROR(IF({eps}="","",{eps}*{eps_factor}),"")',
                    f'=IFERROR({pe}*{pe_factor},"")',
                    f'=IFERROR(B{current_row}*C{current_row},"")',
                ],
                fill=alt_fill if idx % 2 else white_fill,
            )
            ws.cell(current_row, 2).number_format = "$0.00"
            ws.cell(current_row, 3).number_format = "0.0x"
            ws.cell(current_row, 4).number_format = "$0.00"
        row += 1

        row = _section(row, "Adj EBITDA x EV/EBITDA")
        row = _header(row, ["Multiple", "EV", "Equity value: core net cash", "Share price", "Source / investment read"], merge_spans=[(5, max_col)])
        for idx, delta in enumerate((-2.0, 0.0, 2.0)):
            current_row = row
            row = _row(
                row,
                [
                    f'=IFERROR(MAX(0,{ev_multiple}+{delta}),"")',
                    f'=IFERROR(A{current_row}*{ebitda},"")',
                    f'=IFERROR(B{current_row}-{net_debt},"")',
                    f'=IFERROR(C{current_row}/{shares},"")',
                    "Active Adj EBITDA x scenario EV/EBITDA, less active net debt.",
                ],
                fill=alt_fill if idx % 2 else white_fill,
                spans=[(5, max_col)],
            )
            ws.cell(current_row, 1).number_format = "0.0x"
            for cc in (2, 3):
                ws.cell(current_row, cc).number_format = "#,##0.0"
            ws.cell(current_row, 4).number_format = "$0.00"
        row += 1

        row = _section(row, "FCF Yield Implied Equity Value")
        row = _header(row, ["Yield", "Equity value", "Share price", "Source / note"], merge_spans=[(4, max_col)])
        for idx, factor in enumerate((0.80, 1.00, 1.20)):
            current_row = row
            row = _row(
                row,
                [
                    f'=IFERROR(({fcf_yield_rate})*{factor},"")',
                    f'=IFERROR({fcf}/A{current_row},"")',
                    f'=IFERROR(B{current_row}/{shares},"")',
                    "Active FCF capitalized by scenario equity FCF yield.",
                ],
                fill=alt_fill if idx % 2 else white_fill,
                spans=[(4, max_col)],
            )
            ws.cell(current_row, 1).number_format = "0.0%"
            ws.cell(current_row, 2).number_format = "#,##0.0"
            ws.cell(current_row, 3).number_format = "$0.00"
        return row

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(1, 1, f"{ticker_txt} Investment Case")
    ws.cell(1, 1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(1, max_col + 1):
        ws.cell(1, cc).fill = title_fill
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    guidance_basis_label = _guidance_source_contract_label(ticker_txt)
    ws.cell(2, 1, f"Scenario-based analyst handoff sheet; key numbers are backed by History_Q, {guidance_basis_label}, Operating_Drivers and Valuation.")
    ws.cell(2, 1).font = Font(italic=True, size=9, color="52616B")
    ws.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")
    row = 4

    row = _section(row, "Investment Snapshot")
    for idx, rec in enumerate(_records("Investment Snapshot")):
        row = _row(row, [rec.get("metric"), rec.get("display")], fill=callout_fill if idx in {0, 2} else alt_fill, spans=[(2, max_col)])
        note = str(rec.get("source_note") or rec.get("source") or "").strip()
        if note:
            ws.cell(row - 1, 2).comment = Comment(note, "Codex")
    row += 1

    row, manual_refs = _write_manual_inputs(row)
    row, segment_refs = _write_segment_scenario_inputs(row)
    row, bridge_refs = _write_scenario_driver_bridge(row, manual_refs, segment_refs)
    row = _write_market_pricing(row, manual_refs)
    row += 1

    key_debate_text = ""
    for rec in _records("Investment Snapshot"):
        if str(rec.get("metric") or "").strip().lower() == "key debate":
            key_debate_text = str(rec.get("display") or "").strip()
            break
    if key_debate_text:
        row = _section(row, "Key Debate")
        row = _row(row, ["Debate", key_debate_text], fill=callout_fill, spans=[(2, max_col)])
        ws.cell(row - 1, 2).font = Font(bold=True, size=12, color=dark)
        row += 1

    if has_pbi_profile_packs:
        section_order = [
            "Key Debates",
            "Bear / Base / Bull Scenario",
            "What Market Is Pricing",
            "Quality of Earnings",
            "What needs to happen for the stock to work",
            "Turnaround / EBIT Bridge",
            "FCF / Debt Paydown Bridge",
            "Buybacks vs FCF",
            "Current Guide -> Implied Earnings",
            "What Moves EPS",
            "Valuation Sensitivity",
            "Adj EBITDA x EV/EBITDA",
            "FCF Yield Implied Equity Value",
            "Segment Trend / Lapping Risk",
            "Segment Health",
            "Capital Structure / Refinancing Risk",
            "Guidance Beat/Miss Setup",
        ]
    elif has_commodity_ethanol_pack:
        section_order = [
            "Key Debates",
            "Bear / Base / Bull Scenario",
            "What Market Is Pricing",
            "Quality of Earnings",
            "What needs to happen for the stock to work",
            "Ethanol / Crush Margin Bridge",
            "Policy / 45Z / RFS Bridge",
            "Buybacks vs FCF",
            "Current Guide -> Implied Earnings",
            "What Moves EBITDA",
            "What Moves EPS",
            "Valuation Sensitivity",
            "Adj EBITDA x EV/EBITDA",
            "FCF Yield Implied Equity Value",
            "Margin Cycle / Lapping Risk",
            "Ethanol / Policy Health",
            "FCF / Balance Sheet",
            "Guidance Beat/Miss Setup",
        ]
    else:
        section_order = []
    for section in section_order:
        if section in {"What Market Is Pricing", "Adj EBITDA x EV/EBITDA", "FCF Yield Implied Equity Value"}:
            continue
        recs = _records(section)
        if not recs:
            continue
        if section == "Bear / Base / Bull Scenario":
            row = _write_manual_scenarios(row, manual_refs, bridge_refs)
            row += 1
            continue
        if section == "Valuation Sensitivity":
            row = _write_manual_valuation_sensitivity(row, manual_refs, bridge_refs)
            row += 1
            continue
        row = _section(row, section)
        if section == "Key Debates":
            for idx, rec in enumerate(recs):
                title = str(rec.get("metric") or "").strip()
                current_read = str(rec.get("current_read") or "").strip()
                header_text = title
                if current_read:
                    header_text = f"{title} | Current read: {current_read}" if title else f"Current read: {current_read}"
                row = _row(
                    row,
                    [header_text],
                    fill=subheader_fill,
                    spans=[(1, max_col)],
                )
                for cc in range(1, max_col + 1):
                    ws.cell(row - 1, cc).font = Font(bold=True, size=12, color=dark)
                    ws.cell(row - 1, cc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.row_dimensions[row - 1].height = 24
                row = _row(
                    row,
                    [
                        "Bull evidence",
                        rec.get("bull_evidence") or rec.get("display"),
                        "",
                        "",
                        "",
                        "Bear evidence",
                        rec.get("bear_evidence"),
                    ],
                    fill=alt_fill if idx % 2 else white_fill,
                    spans=[(2, 5), (7, 10)],
                )
                ws.cell(row - 1, 1).font = Font(bold=True, size=12, color=dark)
                ws.cell(row - 1, 6).font = Font(bold=True, size=12, color=dark)
                for cc in range(1, max_col + 1):
                    ws.cell(row - 1, cc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.row_dimensions[row - 1].height = 38
                next_proof = str(rec.get("next_proof_point") or "").strip()
                row = _row(
                    row,
                    ["Next proof point", next_proof],
                    fill=alt_fill if idx % 2 else white_fill,
                    spans=[(2, max_col)],
                )
                ws.cell(row - 1, 1).font = Font(bold=True, size=12, color=dark)
                for cc in range(1, max_col + 1):
                    ws.cell(row - 1, cc).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.row_dimensions[row - 1].height = 26
        elif section == "Bear / Base / Bull Scenario":
            # Section header was already written by the common loop; remove it and
            # let the manual-scenario renderer recreate the full interactive block.
            for cc in range(1, max_col + 1):
                ws.cell(row - 1, cc).value = None
                ws.cell(row - 1, cc).fill = PatternFill(fill_type=None)
                ws.cell(row - 1, cc).border = Border()
            row -= 1
            row = _write_manual_scenarios(row, manual_refs, bridge_refs)
        elif section == "What Market Is Pricing":
            row = _header(row, ["Metric", "Current input", "", "Implied metric / read", "", "", "", "Source / note"], merge_spans=[(2, 3), (4, 7), (8, 10)])
            for idx, rec in enumerate(recs):
                row = _row(
                    row,
                    [rec.get("metric"), rec.get("value") or "", "", rec.get("display"), "", "", "", _source_read(rec)],
                    fill=alt_fill if idx % 2 else white_fill,
                    spans=[(2, 3), (4, 7), (8, 10)],
                )
                ws.row_dimensions[row - 1].height = 26
        elif section == "Quality of Earnings":
            row = _header(row, ["Item", "Impact", "", "Cash?", "Recurring?", "Read"], merge_spans=[(2, 3), (6, 10)])
            for idx, rec in enumerate(recs):
                row = _row(
                    row,
                    [
                        rec.get("metric"),
                        rec.get("display"),
                        "",
                        rec.get("cash_flag"),
                        rec.get("recurring_flag"),
                        rec.get("quality_read"),
                    ],
                    fill=alt_fill if idx % 2 else white_fill,
                    spans=[(2, 3), (6, 10)],
                )
                ws.row_dimensions[row - 1].height = 24
        elif section == "Guidance Beat/Miss Setup":
            row = _header(row, ["Metric", "Current guide / actual", "", "", "", "", "", "Current trend / beat-miss risk"], merge_spans=[(2, 7), (8, 10)])
            for idx, rec in enumerate(recs):
                note = rec.get("beat_miss_risk") or rec.get("current_trend") or rec.get("source_note") or rec.get("source")
                row = _row(row, [rec.get("metric"), rec.get("display"), "", "", "", "", "", note], fill=alt_fill if idx % 2 else white_fill, spans=[(2, 7), (8, 10)])
        elif section == "Valuation Sensitivity":
            for cc in range(1, max_col + 1):
                ws.cell(row - 1, cc).value = None
                ws.cell(row - 1, cc).fill = PatternFill(fill_type=None)
                ws.cell(row - 1, cc).border = Border()
            row -= 1
            row = _write_manual_valuation_sensitivity(row, manual_refs, bridge_refs)
        elif section == "Adj EBITDA x EV/EBITDA":
            row = _header(row, ["Multiple", "EV", "Equity value: core net cash", "Share price", "Source / investment read"], merge_spans=[(5, 10)])
            for idx, rec in enumerate(recs):
                display = rec.get("display")
                ev_read = _extract_display_part(display, r"\bEV\s+(\$[0-9,]+(?:\.[0-9]+)?m?)")
                equity_read = _extract_display_part(display, r"\bequity\s+(\$[0-9,]+(?:\.[0-9]+)?m?)")
                share_read = _extract_display_part(display, r"\b(?:equity/share|share)\s+(\$[0-9,]+(?:\.[0-9]+)?)")
                row = _row(
                    row,
                    [rec.get("metric"), ev_read, equity_read, share_read, _source_read(rec)],
                    fill=alt_fill if idx % 2 else white_fill,
                    spans=[(5, 10)],
                )
        elif section == "FCF Yield Implied Equity Value":
            row = _header(row, ["Yield", "Equity value", "Share price", "Source / note"], merge_spans=[(4, 10)])
            for idx, rec in enumerate(recs):
                display = rec.get("display")
                equity_read = _extract_display_part(display, r"\bequity\s+(\$[0-9,]+(?:\.[0-9]+)?m?)")
                share_read = _extract_display_part(display, r"\bshare\s+(\$[0-9,]+(?:\.[0-9]+)?)")
                row = _row(
                    row,
                    [rec.get("metric"), equity_read, share_read, _source_read(rec)],
                    fill=alt_fill if idx % 2 else white_fill,
                    spans=[(4, 10)],
                )
        else:
            row = _header(row, ["Metric", "Value / read", "", "", "", "", "", "Source / investment read"], merge_spans=[(2, 7), (8, 10)])
            for idx, rec in enumerate(recs):
                note = _source_read(rec)
                row = _row(row, [rec.get("metric"), rec.get("display"), "", "", "", "", "", note], fill=alt_fill if idx % 2 else white_fill, spans=[(2, 7), (8, 10)])
        row += 1

    widths = {
        "A": 42,
        "B": 32,
        "C": 30,
        "D": 28,
        "E": 32,
        "F": 24,
        "G": 24,
        "H": 24,
        "I": 22,
        "J": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
