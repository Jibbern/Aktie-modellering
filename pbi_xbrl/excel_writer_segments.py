from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Pattern, Tuple

import pandas as pd
from openpyxl import load_workbook

from .debt_parser import coerce_number


def latest_segment_financials_workbook(segment_financials_dir: Optional[Path]) -> Optional[Path]:
    if segment_financials_dir is None:
        return None
    workbook_candidates = sorted(
        [p for p in segment_financials_dir.glob("*.xlsx") if p.is_file()],
        key=lambda p: (p.stat().st_mtime if p.exists() else 0.0, p.name.lower()),
        reverse=True,
    )
    return workbook_candidates[0] if workbook_candidates else None


def extract_year_from_cell(value_in: Any) -> Optional[int]:
    txt = str(value_in or "").strip()
    if not txt:
        return None
    m_year = re.search(r"\b(20\d{2})\b", txt)
    if not m_year:
        return None
    try:
        return int(m_year.group(1))
    except Exception:
        return None


def extract_quarter_from_cell(value_in: Any) -> Optional[pd.Timestamp]:
    txt = re.sub(r"\s+", " ", str(value_in or "")).strip()
    if not txt or re.search(r"\bfy\b", txt, re.I):
        return None
    mm = re.match(r"^(Mar|Jun|Sep|Dec)\s+(20\d{2})$", txt, re.I)
    if not mm:
        return None
    month_map = {"mar": 3, "jun": 6, "sep": 9, "dec": 12}
    month_num = month_map.get(str(mm.group(1) or "").lower())
    if month_num is None:
        return None
    try:
        return pd.Timestamp(dt.date(int(mm.group(2)), month_num, 1)) + pd.offsets.MonthEnd(0)
    except Exception:
        return None


def extract_segment_line_values(line_in: Any, year_count: int, *, exact_count: bool = False) -> List[float]:
    vals = []
    txt = str(line_in or "")
    for mm in re.finditer(r"(?<!\d)(?:\(?-?[0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?\)?)(?!\d)", txt):
        token = str(mm.group(0) or "").replace("(", "-").replace(")", "").replace(",", "")
        try:
            vals.append(float(token))
        except Exception:
            continue
    if year_count <= 0 or len(vals) < year_count:
        return []
    if exact_count and len(vals) != year_count:
        return []
    return [float(v) * 1000.0 for v in vals[-year_count:]]


def annual_segment_label(
    metric_label: str,
    line_in: Any,
    *,
    annual_segment_alias_patterns: Iterable[Tuple[Pattern[str], str]],
    company_segment_alias_patterns: Iterable[Tuple[Pattern[str], str]],
) -> str:
    txt = re.sub(r"\(\d+\)", "", str(line_in or "")).strip()
    if not txt:
        return ""
    low = txt.lower()
    for pat, label in annual_segment_alias_patterns:
        try:
            if pat.search(low):
                return str(label)
        except Exception:
            continue
    for pat, label in company_segment_alias_patterns:
        try:
            if pat.search(low):
                return str(label)
        except Exception:
            continue
    if metric_label == "Total assets" and re.search(r"\bcorporate\b", low, re.I):
        return "Corporate assets"
    if metric_label in {"Operating income (loss)", "Depreciation & amortization", "Gross margin"} and re.fullmatch(r"corporate", low, re.I):
        return "Corporate expense"
    if re.search(r"\bintersegment eliminations?\b", low, re.I):
        return "Intersegment eliminations"
    if re.search(r"\bother operations?\b", low, re.I):
        return "Other operations"
    return txt


def quarterly_segment_label(
    metric_label: str,
    line_in: Any,
    *,
    annual_segment_alias_patterns: Iterable[Tuple[Pattern[str], str]],
    company_segment_alias_patterns: Iterable[Tuple[Pattern[str], str]],
) -> str:
    txt = re.sub(r"\(\d+\)", "", str(line_in or "")).strip()
    if not txt:
        return ""
    low = txt.lower()
    if low in {
        "total",
        "segment total",
        "adjusted segment ebit",
        "adjusted segment ebitda",
        "pbi adjusted ebit",
        "pbi adjuted ebitda",
        "pbi adjusted ebitda",
    }:
        return ""
    if re.search(r"\btotal\b", low, re.I) and re.search(r"\b(segment|segments|reportable)\b", low, re.I):
        return ""
    if metric_label in {"Adjusted EBIT", "EBIT margin %", "Depreciation & amortization"} and re.fullmatch(r"corporate", low, re.I):
        return "Corporate expense"
    if re.search(r"\bother operations?\b", low, re.I):
        return "Other operations"
    for pat, label in annual_segment_alias_patterns:
        try:
            if pat.search(low):
                return str(label)
        except Exception:
            continue
    for pat, label in company_segment_alias_patterns:
        try:
            if pat.search(low):
                return str(label)
        except Exception:
            continue
    return txt


def parse_quarterly_segment_data_from_workbook(
    path_in: Path,
    *,
    annual_segment_alias_patterns: Iterable[Tuple[Pattern[str], str]],
    company_segment_alias_patterns: Iterable[Tuple[Pattern[str], str]],
) -> Dict[str, Any]:
    out: Dict[str, Any] = {"metrics": {}, "quarters": []}
    try:
        wb_local = load_workbook(path_in, data_only=True, read_only=True)
    except Exception:
        return {}

    def _row_label_local(values_in: List[Any]) -> str:
        for val in values_in[:3]:
            txt = re.sub(r"\s+", " ", str(val or "")).strip()
            if txt:
                return txt
        return ""

    def _quarter_cols_local(ws_local: Any) -> Dict[int, pd.Timestamp]:
        best: Dict[int, pd.Timestamp] = {}
        max_rows = min(int(getattr(ws_local, "max_row", 0) or 0), 16)
        max_cols = min(int(getattr(ws_local, "max_column", 0) or 0), 24)
        for ridx in range(1, max_rows + 1):
            cand: Dict[int, pd.Timestamp] = {}
            for cidx in range(1, max_cols + 1):
                qd = extract_quarter_from_cell(ws_local.cell(ridx, cidx).value)
                if qd is not None:
                    cand[cidx] = pd.Timestamp(qd).normalize()
            if len(cand) >= 2:
                best = cand
                break
        return best

    def _store_segment_values(
        metric_label: str,
        segment_label: str,
        row_vals: List[Any],
        quarter_cols: Dict[int, pd.Timestamp],
        *,
        value_scale: float,
    ) -> None:
        if not metric_label or not segment_label or not quarter_cols:
            return
        q_values: Dict[pd.Timestamp, float] = {}
        for cidx, qd in quarter_cols.items():
            cell_val = row_vals[cidx - 1] if 0 <= (cidx - 1) < len(row_vals) else None
            num_val = coerce_number(cell_val)
            if num_val is None:
                continue
            q_values[pd.Timestamp(qd)] = float(num_val) * float(value_scale)
        if not q_values:
            return
        out["metrics"].setdefault(metric_label, {}).setdefault(segment_label, {}).update(q_values)
        out["quarters"] = sorted(set(out["quarters"]) | set(q_values.keys()))

    if "Revenue & Gross Profit" in wb_local.sheetnames:
        ws_local = wb_local["Revenue & Gross Profit"]
        quarter_cols = _quarter_cols_local(ws_local)
        current_metric = ""
        max_rows = int(getattr(ws_local, "max_row", 0) or 0)
        max_cols = max(max(quarter_cols.keys(), default=1), 12)
        for ridx in range(1, max_rows + 1):
            row_vals = [ws_local.cell(ridx, cidx).value for cidx in range(1, max_cols + 1)]
            row_label = _row_label_local(row_vals)
            row_low = row_label.lower()
            if not row_label:
                continue
            if row_low == "revenue by segment":
                current_metric = "Revenue"
                continue
            if row_low.startswith("gross profit and gross profit margin"):
                current_metric = ""
                continue
            if current_metric != "Revenue":
                continue
            segment_label_val = quarterly_segment_label(
                current_metric,
                row_label,
                annual_segment_alias_patterns=annual_segment_alias_patterns,
                company_segment_alias_patterns=company_segment_alias_patterns,
            )
            if not segment_label_val:
                continue
            _store_segment_values(
                current_metric,
                segment_label_val,
                row_vals,
                quarter_cols,
                value_scale=1e6,
            )

    if "Adj Segment Data" in wb_local.sheetnames:
        ws_local = wb_local["Adj Segment Data"]
        quarter_cols = _quarter_cols_local(ws_local)
        current_metric = ""
        current_segment = ""
        max_rows = int(getattr(ws_local, "max_row", 0) or 0)
        max_cols = max(max(quarter_cols.keys(), default=1), 12)
        for ridx in range(1, max_rows + 1):
            row_vals = [ws_local.cell(ridx, cidx).value for cidx in range(1, max_cols + 1)]
            row_label = _row_label_local(row_vals)
            row_low = row_label.lower()
            if not row_label:
                continue
            if row_low == "adjusted ebit":
                current_metric = "Adjusted EBIT"
                current_segment = ""
                continue
            if row_low == "depreciation & amortization":
                current_metric = "Depreciation & amortization"
                current_segment = ""
                continue
            if row_low in {
                "adjusted ebitda",
                "adjusted segment ebitda",
                "pbi adjusted ebit",
                "pbi adjusted ebitda",
                "pbi adjuted ebitda",
            }:
                current_metric = ""
                current_segment = ""
                continue
            if not current_metric:
                continue
            if current_metric == "Adjusted EBIT" and row_low == "ebit margin":
                if current_segment:
                    _store_segment_values(
                        "EBIT margin %",
                        current_segment,
                        row_vals,
                        quarter_cols,
                        value_scale=1.0,
                    )
                continue
            segment_label_val = quarterly_segment_label(
                current_metric,
                row_label,
                annual_segment_alias_patterns=annual_segment_alias_patterns,
                company_segment_alias_patterns=company_segment_alias_patterns,
            )
            if not segment_label_val:
                current_segment = ""
                continue
            current_segment = segment_label_val if current_metric == "Adjusted EBIT" else current_segment
            _store_segment_values(
                current_metric,
                segment_label_val,
                row_vals,
                quarter_cols,
                value_scale=1e6,
            )

    try:
        wb_local.close()
    except Exception:
        pass
    if out["metrics"]:
        out["source_doc"] = str(path_in)
    return out


def parse_annual_segment_data_from_workbook(
    path_in: Path,
    *,
    annual_segment_alias_patterns: Iterable[Tuple[Pattern[str], str]],
    company_segment_alias_patterns: Iterable[Tuple[Pattern[str], str]],
) -> Dict[str, Any]:
    out: Dict[str, Any] = {"metrics": {}, "assets": {}, "years": []}
    try:
        wb_local = load_workbook(path_in, data_only=True, read_only=True)
    except Exception:
        return {}
    annual_segment_noise_re = re.compile(
        r"\b(ebitda margin|adjusted segment ebitda|pbi adj(?:u)?ted ebitda|"
        r"adjusted segment ebit|pbi adjusted ebit)\b",
        re.I,
    )

    metric_sheet_map = {
        "Revenue & Gross Profit": {
            "Revenue by Segment": "Revenues",
            "Gross Profit and Gross Profit Margin": "Gross margin",
        },
        "Adj Segment Data": {
            "Adjusted EBIT": "Operating income (loss)",
            "Depreciation & amortization": "Depreciation & amortization",
        },
    }

    def _row_label_local(values_in: List[Any]) -> str:
        for val in values_in[:3]:
            txt = re.sub(r"\s+", " ", str(val or "")).strip()
            if txt:
                return txt
        return ""

    def _year_cols_local(ws_local: Any) -> Dict[int, int]:
        best: Dict[int, int] = {}
        max_rows = min(int(getattr(ws_local, "max_row", 0) or 0), 16)
        max_cols = min(int(getattr(ws_local, "max_column", 0) or 0), 18)
        for ridx in range(1, max_rows + 1):
            cand: Dict[int, int] = {}
            for cidx in range(1, max_cols + 1):
                year_val = extract_year_from_cell(ws_local.cell(ridx, cidx).value)
                if year_val is not None:
                    cand[cidx] = year_val
            if len(cand) >= 2:
                best = cand
                break
        return best

    for sheet_name, section_metric_map in metric_sheet_map.items():
        if sheet_name not in wb_local.sheetnames:
            continue
        ws_local = wb_local[sheet_name]
        year_cols = _year_cols_local(ws_local)
        if not year_cols:
            continue
        out["years"] = sorted(set(out["years"]) | set(year_cols.values()))
        current_metric = ""
        max_rows = int(getattr(ws_local, "max_row", 0) or 0)
        max_cols = min(int(getattr(ws_local, "max_column", 0) or 0), max(max(year_cols.keys(), default=1), 12))
        for ridx in range(1, max_rows + 1):
            row_vals = [ws_local.cell(ridx, cidx).value for cidx in range(1, max_cols + 1)]
            row_label = _row_label_local(row_vals)
            row_low = row_label.lower()
            if not row_label:
                continue
            if annual_segment_noise_re.search(row_label):
                continue
            for section_label, metric_label in section_metric_map.items():
                if row_low == section_label.lower():
                    current_metric = metric_label
                    break
            if not current_metric:
                continue
            if current_metric == "Revenues" and row_low == "total":
                continue
            if current_metric == "Gross margin" and row_low in {"gross margin %", "segment gross profit"}:
                continue
            if current_metric == "Operating income (loss)" and row_low in {"ebit margin", "adjusted segment ebit", "pbi adjusted ebit"}:
                continue
            if current_metric == "Depreciation & amortization" and row_low in {"segment total", "total"}:
                continue
            segment_label_val = annual_segment_label(
                current_metric,
                row_label,
                annual_segment_alias_patterns=annual_segment_alias_patterns,
                company_segment_alias_patterns=company_segment_alias_patterns,
            )
            if not segment_label_val:
                continue
            year_values: Dict[int, float] = {}
            for cidx, year_val in year_cols.items():
                cell_val = ws_local.cell(ridx, cidx).value
                num_val = coerce_number(cell_val)
                if num_val is None:
                    continue
                year_values[int(year_val)] = float(num_val) * 1e6
            if not year_values:
                continue
            metric_store = out["assets"] if current_metric == "Total assets" else out["metrics"].setdefault(current_metric, {})
            metric_store.setdefault(segment_label_val, {}).update(year_values)

    try:
        wb_local.close()
    except Exception:
        pass
    if out["metrics"] or out["assets"]:
        out["source_doc"] = str(path_in)
    return out
