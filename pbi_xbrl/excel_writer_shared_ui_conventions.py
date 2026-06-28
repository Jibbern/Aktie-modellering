"""Workbook-wide shared UI convention pass.

This module owns the final visible workbook polish pass that runs after sheet
writers have produced their surfaces.  It is intentionally runtime-backed so
the writer context keeps callback ordering and context-owned helper behavior.
"""
from __future__ import annotations

import math as _math
import re as _re
from copy import copy as _copy
from dataclasses import dataclass
from datetime import date as _date
from typing import Any, Dict, Iterable, List, Mapping, MutableMapping, Optional, Sequence, Set, Tuple

import pandas as _pd
from openpyxl import Workbook
from openpyxl.comments import Comment as _Comment
from openpyxl.styles import Alignment as _Alignment, Border as _Border, Font as _Font, PatternFill as _PatternFill, Side as _Side
from openpyxl.utils import get_column_letter as _get_column_letter


@dataclass(frozen=True)
class SharedUiConventionsDeps:
    runtime: MutableMapping[str, Any]


def apply_shared_ui_conventions_to_workbook(deps: SharedUiConventionsDeps, wb: Any, ticker: Any = "") -> None:
    """Final light-touch polish shared by PBI, GPRE and ANF visible UI sheets."""
    runtime = deps.runtime
    PatternFill = runtime.get("PatternFill", _PatternFill)
    Border = runtime.get("Border", _Border)
    Side = runtime.get("Side", _Side)
    Font = runtime.get("Font", _Font)
    Alignment = runtime.get("Alignment", _Alignment)
    Comment = runtime.get("Comment", _Comment)
    get_column_letter = runtime.get("get_column_letter", _get_column_letter)
    copy = runtime.get("copy", _copy)
    pd = runtime.get("pd", _pd)
    math = runtime.get("math", _math)
    re = runtime.get("re", _re)
    date = runtime.get("date", _date)
    _shared_visible_period_text = runtime["_shared_visible_period_text"]
    _shared_readable_source_label = runtime["_shared_readable_source_label"]
    _standardize_quarter_notes_ui_categories = runtime["_standardize_quarter_notes_ui_categories"]
    _remove_empty_promise_revision_blocks = runtime["_remove_empty_promise_revision_blocks"]
    _polish_promise_scorecard_layout = runtime["_polish_promise_scorecard_layout"]
    _apply_source_backed_promise_mapping_overrides = runtime["_apply_source_backed_promise_mapping_overrides"]
    _polish_investment_case_readability = runtime["_polish_investment_case_readability"]
    _date_or_none = runtime["_date_or_none"]
    _promise_progress_label = runtime["_promise_progress_label"]
    PROMISE_TIMELINE_HEADERS = runtime["PROMISE_TIMELINE_HEADERS"]
    PROMISE_VISIBLE_MAX_COL = runtime["PROMISE_VISIBLE_MAX_COL"]
    ticker_txt = str(ticker or "").strip().upper()
    visible_exact = {
        "SUMMARY",
        "Valuation",
        "BS_Segments",
        "Operating_Drivers",
        "Quarter_Notes_UI",
        "Promise_Progress_UI",
        "ANF_Investment_Case",
        "Economics_Overlay",
    }

    def _is_visible_ui_sheet(name: str) -> bool:
        return (
            name in visible_exact
            or name.endswith("_Investment_Case")
            or name.endswith("_Economics_Overlay")
        )

    def _normalize_cell_text(txt: str) -> str:
        out = _shared_visible_period_text(txt)
        out = _shared_readable_source_label(out)
        return out

    def _latest_history_quarter_label() -> str:
        if "History_Q" not in getattr(wb, "sheetnames", []):
            return "2026-Q1"
        hist_ws = wb["History_Q"]
        if int(hist_ws.max_row or 0) < 2:
            return "2026-Q1"
        headers = [str(hist_ws.cell(1, cc).value or "").strip().lower() for cc in range(1, int(hist_ws.max_column or 0) + 1)]
        if "quarter" not in headers:
            return "2026-Q1"
        q_col = headers.index("quarter") + 1
        latest = None
        for rr in range(2, int(hist_ws.max_row or 0) + 1):
            raw = hist_ws.cell(rr, q_col).value
            try:
                parsed = pd.Timestamp(raw)
            except Exception:
                continue
            if pd.isna(parsed):
                continue
            if latest is None or parsed > latest:
                latest = parsed
        if latest is None:
            return "2026-Q1"
        qn = ((int(latest.month) - 1) // 3) + 1
        return f"{int(latest.year)}-Q{qn}"

    def _fix_gtx_summary_quarter_label() -> None:
        if ticker_txt != "GTX" or "SUMMARY" not in getattr(wb, "sheetnames", []):
            return
        ws = wb["SUMMARY"]
        label = _latest_history_quarter_label()
        for rr in range(1, int(ws.max_row or 0) + 1):
            if str(ws.cell(rr, 1).value or "").strip().lower() == "as of quarter":
                ws.cell(rr, 2).value = label
                ws.cell(rr, 2).number_format = "@"
                ws.cell(rr, 2).alignment = Alignment(horizontal="left", vertical="center")
                return

    def _rewrite_gtx_promise_progress_dashboard() -> None:
        if ticker_txt != "GTX" or "Promise_Progress_UI" not in getattr(wb, "sheetnames", []):
            return
        ws = wb["Promise_Progress_UI"]
        for merged in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged))
        if int(ws.max_row or 0) > 1:
            ws.delete_rows(1, int(ws.max_row or 0))
        title_fill = PatternFill("solid", fgColor="5B9BD5")
        section_fill = PatternFill("solid", fgColor="D9EAF7")
        header_fill = PatternFill("solid", fgColor="EAF3FB")
        thin = Border(
            left=Side(style="thin", color="D9E2EA"),
            right=Side(style="thin", color="D9E2EA"),
            top=Side(style="thin", color="D9E2EA"),
            bottom=Side(style="thin", color="D9E2EA"),
        )

        status_fills = {
            "completed": PatternFill("solid", fgColor="64C2A6"),
            "open": PatternFill("solid", fgColor="99CCFF"),
            "reported": PatternFill("solid", fgColor="99CCFF"),
            "watch": PatternFill("solid", fgColor="F4B183"),
            "raised": PatternFill("solid", fgColor="5B9BD5"),
            "updated": PatternFill("solid", fgColor="5B9BD5"),
            "initial": PatternFill("solid", fgColor="D9EAF7"),
            "post-quarter": PatternFill("solid", fgColor="F4B183"),
        }

        def _status_fill(value: Any) -> Any:
            key = str(value or "").strip().lower()
            return status_fills.get(key, PatternFill("solid", fgColor="D9EAF7"))

        def _style_cell(row_idx: int, col_idx: int, *, fill: Any, bold: bool = False, size: float = 11.0) -> None:
            cell = ws.cell(row_idx, col_idx)
            cell.fill = copy(fill)
            cell.border = thin
            cell.font = Font(bold=bold, color="1F2933", size=size)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        def _merge_write(
            row_idx: int,
            start_col: int,
            end_col: int,
            value: Any,
            *,
            fill: Any,
            bold: bool = False,
            size: float = 11.0,
        ) -> None:
            if end_col > start_col:
                ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
            cell = ws.cell(row_idx, start_col, value)
            for cc in range(start_col, end_col + 1):
                _style_cell(row_idx, cc, fill=fill, bold=bold, size=size)
            cell.value = value

        def _section(row_idx: int, title: str) -> int:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=12)
            cell = ws.cell(row_idx, 1, title)
            cell.fill = title_fill
            cell.border = thin
            cell.font = Font(bold=True, color="FFFFFF", size=13)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            return row_idx + 1

        def _headers(row_idx: int, headers: Sequence[Tuple[int, int, str]]) -> int:
            for start_col, end_col, header in headers:
                _merge_write(row_idx, start_col, end_col, header, fill=header_fill, bold=True, size=11.0)
            return row_idx + 1

        def _row_fill(row_idx: int) -> Any:
            return PatternFill("solid", fgColor="F8FBFD" if row_idx % 2 == 0 else "FFFFFF")

        def _write_open_rows(row_idx: int, records: Iterable[Tuple[Any, ...]]) -> int:
            for metric, guide, horizon, status, note in records:
                fill = _row_fill(row_idx)
                _merge_write(row_idx, 1, 1, metric, fill=fill)
                _merge_write(row_idx, 2, 2, guide, fill=fill)
                _merge_write(row_idx, 3, 3, horizon, fill=fill)
                _merge_write(row_idx, 4, 4, status, fill=_status_fill(status))
                _merge_write(row_idx, 5, 12, note, fill=fill)
                ws.row_dimensions[row_idx].height = 30
                row_idx += 1
            return row_idx

        def _write_actual_rows(row_idx: int, records: Iterable[Tuple[Any, ...]]) -> int:
            for metric, actual, period, status, note in records:
                fill = _row_fill(row_idx)
                _merge_write(row_idx, 1, 1, metric, fill=fill)
                _merge_write(row_idx, 2, 2, actual, fill=fill)
                _merge_write(row_idx, 3, 3, period, fill=fill)
                _merge_write(row_idx, 4, 4, status, fill=_status_fill(status))
                _merge_write(row_idx, 5, 12, note, fill=fill)
                ws.row_dimensions[row_idx].height = 30
                row_idx += 1
            return row_idx

        def _write_context_rows(row_idx: int, records: Iterable[Tuple[Any, ...]]) -> int:
            for period, actuals, progress, status, note in records:
                fill = _row_fill(row_idx)
                _merge_write(row_idx, 1, 1, period, fill=fill)
                _merge_write(row_idx, 2, 4, actuals, fill=fill)
                _merge_write(row_idx, 5, 6, progress, fill=fill)
                _merge_write(row_idx, 7, 7, status, fill=_status_fill(status))
                _merge_write(row_idx, 8, 12, note, fill=fill)
                ws.row_dimensions[row_idx].height = 34
                row_idx += 1
            return row_idx

        def _write_revision_rows(row_idx: int, records: Iterable[Tuple[Any, ...]]) -> int:
            for metric, previous, current, change, actual, progress, status, horizon, stated, note in records:
                fill = _row_fill(row_idx)
                _merge_write(row_idx, 1, 1, metric, fill=fill)
                _merge_write(row_idx, 2, 2, previous, fill=fill)
                _merge_write(row_idx, 3, 4, current, fill=fill)
                _merge_write(row_idx, 5, 5, change, fill=fill)
                _merge_write(row_idx, 6, 6, actual, fill=fill)
                _merge_write(row_idx, 7, 7, progress, fill=fill)
                _merge_write(row_idx, 8, 8, status, fill=_status_fill(status))
                _merge_write(row_idx, 9, 9, horizon, fill=fill)
                _merge_write(row_idx, 10, 10, stated, fill=fill)
                _merge_write(row_idx, 11, 12, note, fill=fill)
                ws.row_dimensions[row_idx].height = 34
                row_idx += 1
            return row_idx

        def _write_event_rows(row_idx: int, records: Iterable[Tuple[Any, ...]]) -> int:
            for event, current, disclosed, status, period, treatment in records:
                fill = _row_fill(row_idx)
                _merge_write(row_idx, 1, 1, event, fill=fill)
                _merge_write(row_idx, 2, 4, current, fill=fill)
                _merge_write(row_idx, 5, 5, disclosed, fill=fill)
                _merge_write(row_idx, 6, 6, status, fill=_status_fill(status))
                _merge_write(row_idx, 7, 8, period, fill=fill)
                _merge_write(row_idx, 9, 12, treatment, fill=fill)
                ws.row_dimensions[row_idx].height = 34
                row_idx += 1
            return row_idx

        ws.sheet_view.zoomScale = 90
        ws.freeze_panes = "A5"
        ws.merge_cells("A1:L1")
        ws["A1"] = "Promise Progress"
        ws["A1"].fill = title_fill
        ws["A1"].border = thin
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("A2:L2")
        ws["A2"] = "GTX guidance dashboard | only clean official guidance revisions are shown"
        ws["A2"].fill = section_fill
        ws["A2"].border = thin
        ws["A2"].font = Font(italic=True, color="1F2933", size=12)
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        rr = 3
        rr = _section(rr, "Management Credibility Scorecard")
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=6)
        ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=12)
        for cc, header in ((1, "Category"), (2, "Score"), (3, "Evidence"), (7, "Read")):
            cell = ws.cell(rr, cc, header)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="1F2933", size=11)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for cc in range(1, 13):
            ws.cell(rr, cc).fill = header_fill
            ws.cell(rr, cc).border = thin
        rr += 1
        for record in (
            ("2026 outlook", "Open", "Raised FY2026 outlook after Q1; Q1 net sales $985m, adjusted EBIT $151m and adjusted FCF $49m.", "Guidance remains open; Q1 is a read-through, not full-year completion."),
            ("Capital allocation", "Watch", "Q1 buybacks $87m; remaining authorization $163m; FY2025 buybacks were $208m.", "Watch leverage, share count and unrestricted cash before giving buybacks full valuation credit."),
            ("Debt event separation", "Watch", "Post-quarter May 18 2026 $50m term-loan repayment/repricing disclosed.", "Event context only; do not rewrite Q1 reported history."),
        ):
            ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=6)
            ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=12)
            row_fill = PatternFill("solid", fgColor="F8FBFD" if rr % 2 == 0 else "FFFFFF")
            ws.cell(rr, 1, record[0])
            ws.cell(rr, 2, record[1])
            ws.cell(rr, 3, record[2])
            ws.cell(rr, 7, record[3])
            for cc in range(1, 13):
                cell = ws.cell(rr, cc)
                cell.fill = row_fill
                cell.border = thin
                cell.font = Font(color="1F2933", size=11)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.cell(rr, 2).fill = _status_fill(record[1])
            ws.row_dimensions[rr].height = 34
            rr += 1
        rr += 1
        rr = _section(rr, "2026 open guidance")
        rr = _headers(rr, [(1, 1, "Metric"), (2, 2, "Current guide"), (3, 3, "Horizon"), (4, 4, "Status"), (5, 12, "Notes/source")])
        rr = _write_open_rows(
            rr,
            [
                ("Net sales", "$3.6bn-$3.9bn", "2026 year", "Open", "Raised after Q1; Q1 actual was $985m. Source: Q1 2026 earnings release."),
                ("Constant-currency sales growth", "-2% to +6%", "2026 year", "Open", "Updated range; demand backdrop stays industry-sensitive. Source: Q1 2026 earnings release."),
                ("Net income", "$300m-$360m", "2026 year", "Open", "Raised GAAP guide; Q1 net income was $95m. Source: Q1 2026 earnings release."),
                ("Adjusted EBIT", "$520m-$600m", "2026 year", "Open", "Primary non-GAAP guide; Q1 adjusted EBIT was $151m. Source: Q1 2026 earnings release."),
                ("CFO", "$407m-$522m", "2026 year", "Open", "GAAP cash-flow guide; Q1 CFO was $98m. Source: Q1 2026 earnings release."),
                ("Adjusted FCF", "$355m-$475m", "2026 year", "Open", "Company-defined adjusted FCF; Q1 adjusted FCF was $49m. Source: Q1 2026 earnings release."),
                ("Light vehicle production", "down 1%-3%", "2026 year", "Watch", "End-market assumption. Source: Q1 2026 earnings release."),
                ("Commercial vehicle industry", "up 1%-2%", "2026 year", "Watch", "End-market assumption. Source: Q1 2026 earnings release."),
                ("BEV penetration", "~19%", "2026 year", "Watch", "Powertrain mix assumption. Source: Q1 2026 earnings release."),
                ("EUR/USD", "1.17", "2026 year", "Watch", "FX assumption. Source: Q1 2026 earnings release."),
                ("RD&E", "4.2% of sales", "2026 year", "Watch", "Technology investment assumption; keep as outlook/driver, not historical RD&E series. Source: Q1 2026 earnings release."),
                ("Capex", "2.5% of sales", "2026 year", "Watch", "FCF bridge assumption. Source: Q1 2026 earnings release."),
            ],
        )
        rr += 1
        rr = _section(rr, "2025 completed actuals")
        rr = _headers(rr, [(1, 1, "Metric"), (2, 2, "Actual"), (3, 3, "Period"), (4, 4, "Status"), (5, 12, "Notes/source")])
        rr = _write_actual_rows(
            rr,
            [
                ("Q4 2025 net sales", "$891m", "2025-Q4", "Completed", "Quarter value, not FY. Source: Q4 2025 earnings release."),
                ("Q4 2025 adjusted EBIT", "$122m", "2025-Q4", "Completed", "Quarter value. Source: Q4 2025 earnings release."),
                ("Q4 2025 adjusted FCF", "$139m", "2025-Q4", "Completed", "Quarter value. Source: Q4 2025 earnings release."),
                ("FY2025 net sales", "$3.584bn", "FY2025", "Completed", "Completed annual actual. Source: Q4 2025 earnings release."),
                ("FY2025 adjusted EBIT", "$510m", "FY2025", "Completed", "Completed annual actual. Source: Q4 2025 earnings release."),
                ("FY2025 adjusted FCF", "$403m", "FY2025", "Completed", "Completed annual actual. Source: Q4 2025 earnings release."),
                ("FY2025 buybacks", "$208m", "FY2025", "Completed", "Capital allocation actual; common share count reduction 8% year-over-year. Source: Q4 2025 earnings release."),
            ],
        )
        rr += 1
        rr = _section(rr, "Quarterly actual/context rows")
        rr = _headers(rr, [(1, 1, "Quarter"), (2, 4, "Actuals"), (5, 6, "Progress / run-rate"), (7, 7, "Status"), (8, 12, "Notes/source")])
        rr = _write_context_rows(
            rr,
            [
                ("2026-Q1", "Sales $985m; adj EBIT $151m; adj EBITDA $183m; adj FCF $49m", "CFO $98m; capex $29m; buybacks $87m", "Reported", "High-confidence official rows; post-quarter debt event stays separate."),
                ("2025-Q4", "Sales $891m; adj EBIT $122m; adj EBITDA $159m; adj FCF $139m", "CFO $99m; capex $21m; buybacks $72m", "Reported", "Quarter values; do not select FY columns."),
                ("2025-Q3", "Sales $902m; adj EBIT $133m; adj EBITDA $164m; adj FCF $107m", "CFO $100m; capex $10m; buybacks $84m", "Reported", "Actuals-only context, not a guidance revision."),
                ("2025-Q2", "Sales $913m; adj EBIT $124m; adj EBITDA $154m; adj FCF $121m", "CFO $158m; capex $15m; buybacks $22m", "Reported", "Actuals-only context, not a guidance revision."),
                ("2025-Q1", "Sales $878m; adj EBIT $131m; adj EBITDA $159m; adj FCF $36m", "CFO $56m; capex $26m; buybacks $30m", "Reported", "Actuals-only context, not a guidance revision."),
            ],
        )
        rr += 1
        rr = _section(rr, "Quarterly guidance timeline / revision log")
        rr = _headers(rr, [(1, 1, "Metric"), (2, 2, "Previous guide"), (3, 4, "New/current guide"), (5, 5, "Change type"), (6, 6, "Actual"), (7, 7, "Progress / run-rate"), (8, 8, "Status"), (9, 9, "Horizon"), (10, 10, "Stated in"), (11, 12, "Notes/source")])
        rr = _write_revision_rows(
            rr,
            [
                ("Initial FY2026 outlook", "", "FY2026 outlook issued with FY2025 results", "Initial", "FY2025 actuals completed", "Base guide from Q4/FY2025 release", "Open", "2026 year", "2025-Q4", "Q4 2025 earnings release; track revisions against Q1 guide."),
                ("Raised FY2026 outlook", "Initial 2026 outlook", "Net sales $3.6bn-$3.9bn; adj EBIT $520m-$600m; adj FCF $355m-$475m", "Raised", "Q1 2026 actuals reported", "Annual guide remains open", "Open", "2026 year", "2026-Q1", "Q1 2026 earnings release."),
            ],
        )
        rr += 1
        rr = _section(rr, "Post-quarter May 2026 debt repayment/repricing event")
        rr = _headers(rr, [(1, 1, "Event"), (2, 4, "Current/pro-forma event"), (5, 5, "Disclosed value"), (6, 6, "Status"), (7, 8, "Period/event"), (9, 12, "Treatment/source")])
        rr = _write_event_rows(
            rr,
            [
                ("Debt reduction", "Post-quarter event-context / pro-forma: $50m term-loan repayment/repricing", "$50m disclosed post-quarter", "Watch", "Post-quarter May 18 2026", "May 18 2026 press release / 8-K package; not Q1 reported history."),
            ],
        )

        widths = {1: 28, 2: 24, 3: 22, 4: 18, 5: 20, 6: 20, 7: 18, 8: 18, 9: 18, 10: 18, 11: 28, 12: 28, 13: 3, 14: 3, 15: 3}
        for cc, width in widths.items():
            ws.column_dimensions[get_column_letter(cc)].width = width
            if cc >= 13:
                ws.column_dimensions[get_column_letter(cc)].hidden = True
        for row_idx in range(1, int(ws.max_row or 0) + 1):
            if row_idx in {1, 2}:
                ws.row_dimensions[row_idx].height = 28
            elif str(ws.cell(row_idx, 1).value or "").strip() in {
                "Management Credibility Scorecard",
                "2026 open guidance",
                "2025 completed actuals",
                "Quarterly actual/context rows",
                "Quarterly guidance timeline / revision log",
                "Post-quarter May 2026 debt repayment/repricing event",
            }:
                ws.row_dimensions[row_idx].height = 26
            else:
                ws.row_dimensions[row_idx].height = max(float(ws.row_dimensions[row_idx].height or 0), 30.0 if row_idx >= 5 else 24.0)

    def _rewrite_gtx_valuation_guidance_and_driver_panels() -> None:
        if ticker_txt != "GTX" or "Valuation" not in getattr(wb, "sheetnames", []):
            return
        ws = wb["Valuation"]
        panel_start = 15  # O
        panel_end = 28  # AB; keep AC and beyond for existing workbook internals.
        thin = Border(
            left=Side(style="thin", color="D9E2EA"),
            right=Side(style="thin", color="D9E2EA"),
            top=Side(style="thin", color="D9E2EA"),
            bottom=Side(style="thin", color="D9E2EA"),
        )
        title_fill = PatternFill("solid", fgColor="D9EAF7")
        header_fill = PatternFill("solid", fgColor="EAF3FB")
        alt_fill = PatternFill("solid", fgColor="F8FBFD")
        body_fill = PatternFill("solid", fgColor="FFFFFF")

        def _unmerge_panel_rows(first_row: int, last_row: int) -> None:
            for merged in list(ws.merged_cells.ranges):
                if (
                    merged.max_row >= first_row
                    and merged.min_row <= last_row
                    and merged.max_col >= panel_start
                    and merged.min_col <= panel_end
                ):
                    ws.unmerge_cells(str(merged))

        def _clear_panel(first_row: int, last_row: int) -> None:
            _unmerge_panel_rows(first_row, last_row)
            for rr in range(first_row, last_row + 1):
                for cc in range(panel_start, panel_end + 1):
                    cell = ws.cell(rr, cc)
                    cell.value = None
                    cell.comment = None
                    cell.fill = body_fill
                    cell.border = thin
                    cell.font = Font(color="1F2933", size=10)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                ws.row_dimensions[rr].height = 19.5

        def _merge_title(row_idx: int, text: str) -> None:
            ws.merge_cells(start_row=row_idx, start_column=panel_start, end_row=row_idx, end_column=panel_end)
            cell = ws.cell(row_idx, panel_start, text)
            cell.fill = title_fill
            cell.border = thin
            cell.font = Font(bold=True, color="1F2933", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        def _style_row(row_idx: int, *, fill: Any = body_fill, bold: bool = False) -> None:
            for cc in range(panel_start, panel_end + 1):
                cell = ws.cell(row_idx, cc)
                cell.fill = copy(fill)
                cell.border = thin
                cell.font = Font(bold=bold, color="1F2933", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        def _guidance_number(metric: str, stated_in: str) -> str:
            if "Guidance_Normalized" not in getattr(wb, "sheetnames", []):
                return ""
            gws = wb["Guidance_Normalized"]
            headers = [
                str(gws.cell(1, cc).value or "").strip().lower()
                for cc in range(1, int(gws.max_column or 0) + 1)
            ]
            try:
                metric_col = headers.index("metric") + 1
                numbers_col = headers.index("numbers") + 1
                stated_col = headers.index("stated_in_label") + 1
            except ValueError:
                return ""
            metric_low = metric.strip().lower()
            for rr in range(2, int(gws.max_row or 0) + 1):
                if str(gws.cell(rr, stated_col).value or "").strip() != stated_in:
                    continue
                if str(gws.cell(rr, metric_col).value or "").strip().lower() == metric_low:
                    return str(gws.cell(rr, numbers_col).value or "").strip()
            return ""

        def _display(metric: str, stated_in: str, fallback: str) -> str:
            raw = _guidance_number(metric, stated_in)
            if metric == "Light vehicle production" and raw:
                return "down 1%-3%"
            if metric == "Commercial vehicle industry" and raw:
                return "up 1%-2%"
            return raw or fallback

        _clear_panel(7, 18)
        _merge_title(7, "Guidance (As of 2026-03-31) - Status: Raised")
        row_values = {
            8: ["Metric", "Guidance", "Metric", "Guidance", "Metric", "Guidance"],
            9: [
                "Net sales",
                _display("Net sales", "2026-Q1", "$3.6bn-$3.9bn"),
                "Constant-currency sales growth",
                _display("Constant-currency sales growth", "2026-Q1", "-2% to +6%"),
                "Net income",
                _display("Net income", "2026-Q1", "$300m-$360m"),
            ],
            10: [
                "Adjusted EBIT",
                _display("Adjusted EBIT", "2026-Q1", "$520m-$600m"),
                "CFO",
                _display("CFO", "2026-Q1", "$407m-$522m"),
                "Adjusted FCF",
                _display("Adjusted FCF", "2026-Q1", "$355m-$475m"),
            ],
            11: [
                "Light vehicle production",
                _display("Light vehicle production", "2026-Q1", "down 1%-3%"),
                "Commercial vehicle industry",
                _display("Commercial vehicle industry", "2026-Q1", "up 1%-2%"),
                "BEV penetration",
                _display("BEV penetration", "2026-Q1", "about 19%"),
            ],
            12: [
                "EUR/USD",
                _display("EUR/USD", "2026-Q1", "1.17"),
                "RD&E",
                _display("RD&E", "2026-Q1", "about 4.2% of sales"),
                "Capex",
                _display("Capex", "2026-Q1", "about 2.5% of sales"),
            ],
        }
        for rr, values in row_values.items():
            _style_row(rr, fill=header_fill if rr == 8 else (alt_fill if rr % 2 else body_fill), bold=rr == 8)
            for idx, value in enumerate(values):
                ws.cell(rr, panel_start + idx, value)
        _merge_title(13, "Guidance (As of 2025-12-31) - Status: Initial / FY2025 actuals")
        _style_row(14, fill=alt_fill)
        ws.cell(14, panel_start, "FY2025 actuals")
        ws.cell(14, panel_start + 1, _display("FY2025 net sales", "2025-Q4", "$3.584bn"))
        ws.cell(14, panel_start + 2, "FY2025 adjusted EBIT")
        ws.cell(14, panel_start + 3, _display("FY2025 adjusted EBIT", "2025-Q4", "$510m"))
        ws.cell(14, panel_start + 4, "FY2025 adjusted FCF")
        ws.cell(14, panel_start + 5, _display("FY2025 adjusted FCF", "2025-Q4", "$403m"))
        ws.cell(14, panel_start + 6, "FY2025 buybacks")
        ws.cell(14, panel_start + 7, _display("FY2025 buybacks", "2025-Q4", "$208m; 8%"))

        _merge_title(16, "Operating Drivers")
        _style_row(17, fill=header_fill, bold=True)
        for idx, label in enumerate(
            [
                "OEM production / end-market demand",
                "Product mix / turbo demand",
                "Commercial vehicle / industrial",
                "Aftermarket",
                "China / Europe exposure",
                "Customer concentration",
                "RD&E / technology awards",
                "Adjusted EBIT / adjusted FCF conversion",
                "Debt, net leverage and buybacks",
            ]
        ):
            ws.cell(17, panel_start + idx, label)
        _style_row(18, fill=alt_fill)
        ws.cell(18, panel_start, "Compact map; detailed source-backed data remains on Operating_Drivers.")

    def _rewrite_gtx_bs_segments_analytical_cuts() -> None:
        if ticker_txt != "GTX" or "BS_Segments" not in getattr(wb, "sheetnames", []):
            return
        ws = wb["BS_Segments"]
        existing_start: Optional[int] = None
        for rr in range(1, int(ws.max_row or 0) + 1):
            row_text = " | ".join(str(ws.cell(rr, cc).value or "") for cc in range(1, int(ws.max_column or 0) + 1))
            if "Annual analytical cuts" in row_text:
                existing_start = rr
                break
        if existing_start is not None:
            ws.delete_rows(existing_start, int(ws.max_row or 0) - existing_start + 1)
        for rr in range(int(ws.max_row or 0), 0, -1):
            row_text = " | ".join(str(ws.cell(rr, cc).value or "") for cc in range(1, int(ws.max_column or 0) + 1))
            if "No annual segment data found for GTX" in row_text:
                ws.delete_rows(rr, 1)
        title_fill = PatternFill("solid", fgColor="D9EAF7")
        header_fill = PatternFill("solid", fgColor="EAF3FB")
        alt_fill = PatternFill("solid", fgColor="F8FBFD")
        thin = Border(
            left=Side(style="thin", color="D9E2EA"),
            right=Side(style="thin", color="D9E2EA"),
            top=Side(style="thin", color="D9E2EA"),
            bottom=Side(style="thin", color="D9E2EA"),
        )
        rr = int(ws.max_row or 0) + 2
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=9)
        ws.cell(rr, 1, "Annual analytical cuts — not reportable segments")
        for cc in range(1, 10):
            cell = ws.cell(rr, cc)
            cell.fill = title_fill
            cell.border = thin
            cell.font = Font(bold=True, color="1F2933", size=12)
        rr += 1
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=9)
        ws.cell(rr, 1, "GTX reports one operating/reportable segment; these rows are analytical revenue cuts, not segment profit.")
        for cc in range(1, 10):
            cell = ws.cell(rr, cc)
            cell.fill = alt_fill
            cell.border = thin
            cell.font = Font(italic=True, color="1F2933", size=12)
        rr += 2

        def _write_table(title: str, headers_in: Sequence[str], records: Sequence[Sequence[Any]]) -> None:
            nonlocal rr
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=9)
            ws.cell(rr, 1, title)
            for cc in range(1, 10):
                cell = ws.cell(rr, cc)
                cell.fill = title_fill
                cell.border = thin
                cell.font = Font(bold=True, color="1F2933", size=12)
            rr += 1
            display_headers = [*(str(header) for header in headers_in[:4]), "Source / treatment"]
            for cc, header in enumerate(display_headers, start=1):
                cell = ws.cell(rr, cc, header)
                cell.fill = header_fill
                cell.border = thin
                cell.font = Font(bold=True, color="1F2933", size=12)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            for cc in range(len(display_headers) + 1, 10):
                cell = ws.cell(rr, cc)
                cell.fill = header_fill
                cell.border = thin
                cell.font = Font(bold=True, color="1F2933", size=12)
            ws.merge_cells(start_row=rr, start_column=5, end_row=rr, end_column=9)
            rr += 1
            for record in records:
                values = list(record)
                while len(values) < 6:
                    values.append("")
                source_note = " — ".join(str(value or "").strip() for value in values[4:6] if str(value or "").strip())
                display_values = [*values[:4], source_note]
                for cc, value in enumerate(display_values, start=1):
                    cell = ws.cell(rr, cc, value)
                    cell.fill = alt_fill if rr % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                    cell.border = thin
                    cell.font = Font(color="1F2933", size=12)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=cc >= 5)
                for cc in range(len(display_values) + 1, 10):
                    cell = ws.cell(rr, cc)
                    cell.fill = alt_fill if rr % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                    cell.border = thin
                    cell.font = Font(color="1F2933", size=12)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws.merge_cells(start_row=rr, start_column=5, end_row=rr, end_column=9)
                ws.row_dimensions[rr].height = 24 if any(str(value or "").strip() for value in record) else 16
                rr += 1
            rr += 1

        _write_table(
            "Product-line revenue by year",
            ["Product line", "FY2023", "FY2024", "FY2025", "Source", "Treatment"],
            [
                ("Gas ($m)", 1720, 1505, 1592, "2025 Form 10-K", "Product revenue cut"),
                ("Diesel ($m)", 992, 827, 837, "2025 Form 10-K", "Product revenue cut"),
                ("Commercial Vehicles / Industrial ($m)", 656, 629, 654, "2025 Form 10-K", "Product revenue cut"),
                ("Aftermarket ($m)", 456, 459, 438, "2025 Form 10-K", "Product revenue cut"),
                ("Other ($m)", 62, 55, 63, "2025 Form 10-K", "Product revenue cut"),
            ],
        )
        _write_table(
            "Geography revenue by year",
            ["Geography", "FY2023", "FY2024", "FY2025", "Source", "Treatment"],
            [
                ("United States ($m)", 744, 700, 694, "2025 Form 10-K", "Geography revenue cut"),
                ("Europe ($m)", 1874, 1642, 1745, "2025 Form 10-K", "Geography revenue cut"),
                ("China ($m)", 768, 643, 638, "2025 Form 10-K", "Geography revenue cut"),
                ("Rest of Asia ($m)", 433, 413, 406, "2025 Form 10-K", "Geography revenue cut"),
                ("Other International ($m)", 67, 77, 101, "2025 Form 10-K", "Geography revenue cut"),
            ],
        )
        _write_table(
            "Customer concentration",
            ["Customer / group", "FY2023", "FY2024", "FY2025", "Source", "Treatment"],
            [
                ("Stellantis revenue ($m)", 347, 330, 424, "2025 Form 10-K", "Customer watch"),
                ("BMW revenue ($m)", 474, 401, 385, "2025 Form 10-K", "Customer watch"),
                ("Ford revenue ($m)", 364, 354, 377, "2025 Form 10-K", "Customer watch"),
                ("", "", "", "", "", ""),
                ("Stellantis % sales", "9%", "9%", "12%", "2025 Form 10-K", "Customer watch"),
                ("BMW % sales", "12%", "12%", "11%", "2025 Form 10-K", "Customer watch"),
                ("Ford % sales", "9%", "10%", "11%", "2025 Form 10-K", "Customer watch"),
                ("Top ten customers % sales", "", "", "about 62%", "2025 Form 10-K", "Not a segment; customer risk"),
            ],
        )
        for cc, width in {1: 34, 2: 13, 3: 13, 4: 13, 5: 13, 6: 13, 7: 13, 8: 13, 9: 13}.items():
            ws.column_dimensions[get_column_letter(cc)].width = width

    def _clean_gtx_quarter_notes_visible_noise() -> None:
        if ticker_txt != "GTX" or "Quarter_Notes_UI" not in getattr(wb, "sheetnames", []):
            return
        ws = wb["Quarter_Notes_UI"]
        for merged in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged))
        if int(ws.max_row or 0) > 1:
            ws.delete_rows(1, int(ws.max_row or 0))

        title_fill = PatternFill("solid", fgColor="5B9BD5")
        sub_fill = PatternFill("solid", fgColor="DDEBF7")
        table_header_fill = PatternFill("solid", fgColor="EAF3F8")
        body_fill = PatternFill("solid", fgColor="F7FBFF")
        alt_body_fill = PatternFill("solid", fgColor="EDF4FB")
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        thin = Border(
            left=Side(style="thin", color="D9E2EA"),
            right=Side(style="thin", color="D9E2EA"),
            top=Side(style="thin", color="D9E2EA"),
            bottom=Side(style="thin", color="D9E2EA"),
        )

        def _merge(row_idx: int, start_col: int, end_col: int, value: str, *, fill: Any, bold: bool = False, size: float = 11.0) -> None:
            ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
            cell = ws.cell(row_idx, start_col, value)
            cell.fill = fill
            cell.border = thin
            cell.font = Font(bold=bold, color="FFFFFF" if fill == title_fill else "1F2933", size=size)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            for cc in range(start_col, end_col + 1):
                ws.cell(row_idx, cc).fill = fill
                ws.cell(row_idx, cc).border = thin

        def _set_height(row_idx: int, height: float) -> None:
            ws.row_dimensions[row_idx].height = height

        def _spacer(row_idx: int) -> None:
            _set_height(row_idx, 10.0)

        def _label_row(row_idx: int, label: str, text: str, *, fill: Any = body_fill, height: float = 44.0) -> None:
            ws.cell(row_idx, 1, label)
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=15)
            ws.cell(row_idx, 2, text)
            for cc in range(1, 16):
                cell = ws.cell(row_idx, cc)
                cell.fill = fill
                cell.border = thin
                cell.font = Font(bold=cc == 1, color="1F2933", size=11)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            _set_height(row_idx, height)

        def _table_header(row_idx: int) -> None:
            headers = {
                1: "Theme",
                3: "What happened",
                6: "Why it matters",
                8: "Model / valuation implication",
                13: "Source / confidence",
            }
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=5)
            ws.merge_cells(start_row=row_idx, start_column=6, end_row=row_idx, end_column=7)
            ws.merge_cells(start_row=row_idx, start_column=8, end_row=row_idx, end_column=12)
            ws.merge_cells(start_row=row_idx, start_column=13, end_row=row_idx, end_column=15)
            for cc in range(1, 16):
                cell = ws.cell(row_idx, cc)
                if cc in headers:
                    cell.value = headers[cc]
                cell.fill = table_header_fill
                cell.border = thin
                cell.font = Font(bold=True, color="1F2933", size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            _set_height(row_idx, 26.0)

        def _table_row(row_idx: int, theme: str, happened: str, why: str, treatment: str, source: str) -> None:
            ws.cell(row_idx, 1, theme)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=5)
            ws.cell(row_idx, 3, happened)
            ws.merge_cells(start_row=row_idx, start_column=6, end_row=row_idx, end_column=7)
            ws.cell(row_idx, 6, why)
            ws.merge_cells(start_row=row_idx, start_column=8, end_row=row_idx, end_column=12)
            ws.cell(row_idx, 8, treatment)
            ws.merge_cells(start_row=row_idx, start_column=13, end_row=row_idx, end_column=15)
            ws.cell(row_idx, 13, source)
            row_fill = body_fill if row_idx % 2 == 0 else white_fill
            for cc in range(1, 16):
                cell = ws.cell(row_idx, cc)
                cell.fill = row_fill
                cell.border = thin
                cell.font = Font(color="1F2933", size=11)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            _set_height(row_idx, 48.0)

        def _promise_header(row_idx: int) -> None:
            headers = {
                1: "Promise / guidance item",
                3: "Read",
                7: "Actual / progress interpretation",
                13: "Status / source",
            }
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=6)
            ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=12)
            ws.merge_cells(start_row=row_idx, start_column=13, end_row=row_idx, end_column=15)
            for cc in range(1, 16):
                cell = ws.cell(row_idx, cc)
                if cc in headers:
                    cell.value = headers[cc]
                cell.fill = sub_fill
                cell.border = thin
                cell.font = Font(bold=True, color="1F2933", size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            _set_height(row_idx, 26.0)

        def _promise_row(row_idx: int, item: str, read: str, progress: str, status: str, source: str) -> None:
            ws.cell(row_idx, 1, item)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=6)
            ws.cell(row_idx, 3, read)
            ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=12)
            ws.cell(row_idx, 7, progress)
            ws.merge_cells(start_row=row_idx, start_column=13, end_row=row_idx, end_column=15)
            ws.cell(row_idx, 13, " — ".join(part for part in (status, source) if str(part or "").strip()))
            row_fill = body_fill if row_idx % 2 == 0 else white_fill
            for cc in range(1, 16):
                cell = ws.cell(row_idx, cc)
                cell.fill = row_fill
                cell.border = thin
                cell.font = Font(color="1F2933", size=11)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            _set_height(row_idx, 48.0 if max(len(str(read or "")), len(str(progress or "")), len(str(status or ""))) > 115 else 34.0)

        def _quarter_block(row_idx: int, title: str, read: str, changed: str, watch: str, caveat: str, rows_in: Sequence[Tuple[str, str, str, str, str]]) -> int:
            _merge(row_idx, 1, 15, title, fill=title_fill, bold=True, size=14)
            _set_height(row_idx, 24.0)
            row_idx += 1
            _merge(row_idx, 1, 15, "Quarter read", fill=sub_fill, bold=True, size=13)
            _set_height(row_idx, 25.0)
            row_idx += 1
            _label_row(row_idx, "Model read", read, fill=body_fill, height=44.0)
            row_idx += 1
            _label_row(row_idx, "What changed", changed, fill=alt_body_fill, height=44.0)
            row_idx += 1
            _label_row(row_idx, "Watch next", watch, fill=body_fill, height=30.0)
            row_idx += 1
            _label_row(row_idx, "Key caveat", caveat, fill=alt_body_fill, height=30.0)
            row_idx += 1
            _spacer(row_idx)
            row_idx += 1
            _merge(row_idx, 1, 15, "Key developments", fill=sub_fill, bold=True, size=13)
            _set_height(row_idx, 25.0)
            row_idx += 1
            _table_header(row_idx)
            row_idx += 1
            for record in rows_in:
                _table_row(row_idx, *record)
                row_idx += 1
            _spacer(row_idx)
            row_idx += 1
            _merge(row_idx, 1, 15, "Guidance / Promise interpretation", fill=sub_fill, bold=True, size=13)
            _set_height(row_idx, 25.0)
            row_idx += 1
            _promise_header(row_idx)
            row_idx += 1
            _promise_row(
                row_idx,
                "Clean official guidance",
                "Show guidance only when the official release or filing gives a clean row.",
                "Q1 2026 and Q4/FY2025 rows feed Promise_Progress_UI.",
                "No noisy slide OCR or legal boilerplate.",
                "Workbook source package / earnings releases",
            )
            row_idx += 1
            _promise_row(
                row_idx,
                "Actuals-only context",
                "Quarter actuals can explain progress even when no clean guide revision exists.",
                "Use History_Q and high-confidence Adjusted_Metrics rows.",
                "Do not invent revisions for quarters without source-backed guide updates.",
                "History_Q / Adjusted_Metrics",
            )
            row_idx += 1
            _spacer(row_idx)
            row_idx += 1
            _merge(row_idx, 1, 15, "Model mapping / double-count guardrails", fill=sub_fill, bold=True, size=13)
            _set_height(row_idx, 25.0)
            row_idx += 1
            _promise_header(row_idx)
            row_idx += 1
            _promise_row(
                row_idx,
                "Revenue / mix",
                "Product, geography and customer data are analytical cuts.",
                "Use Operating_Drivers and BS_Segments for read-through, not accounting segments.",
                "Keep one reportable-segment treatment.",
                "Official filing tables",
            )
            row_idx += 1
            _promise_row(
                row_idx,
                "Adjusted metrics",
                "Adjusted EBIT, adjusted EBITDA and adjusted FCF are non-GAAP scorecard rows.",
                "Show quarterly values on Valuation, with GAAP FCF kept separate.",
                "Do not overwrite GAAP EBITDA when only adjusted EBITDA is reported.",
                "Earnings releases / Adjusted_Metrics",
            )
            row_idx += 1
            _promise_row(
                row_idx,
                "Special events",
                "May 2026 debt event is post-quarter context.",
                "Keep Q1 reported history unchanged.",
                "Only use as pro-forma/event context.",
                "May 18 2026 release / 8-K package",
            )
            row_idx += 1
            _spacer(row_idx)
            return row_idx + 1

        rr = 1
        rr = _quarter_block(
            rr,
            "2026-Q1 - Quarter Notes",
            "Q1 read: net sales $985m, adjusted EBIT $151m, adjusted EBITDA $183m, adjusted FCF $49m and buybacks $87m.",
            "2026 guidance raised after Q1; management also cited business award highlights across turbo, range-extended EV, E-Powertrain and E-Cooling.",
            "Watch conversion of the raised annual guide into adjusted EBIT, adjusted FCF and unrestricted-cash-backed leverage reduction.",
            "May 18 2026 term-loan repayment/repricing is post-quarter/event context only and does not rewrite Q1 reported history.",
            [
                ("Revenue / guide", "Net sales $985m; 2026 guidance raised to $3.6bn-$3.9bn.", "Sets the near-term revenue base for the annual outlook.", "Use in Valuation guidance and Promise_Progress_UI; keep annual guide distinct from quarter actuals.", "High; Q1 2026 earnings release"),
                ("Adjusted EBIT", "Adjusted EBIT $151m; adjusted EBITDA $183m.", "Primary non-GAAP operating scorecard for GTX.", "Use Adjusted EBIT as primary operating metric and Adjusted EBITDA as secondary bridge.", "High; Q1 2026 earnings release"),
                ("Adjusted FCF", "Adjusted FCF $49m; GAAP CFO minus capex was $69m.", "Shows company-defined cash conversion versus GAAP FCF basis.", "Classify as definition difference, not conflicting extraction.", "High; Q1 2026 earnings release / History_Q"),
                ("Buybacks / leverage", "Buybacks $87m; unrestricted cash $142m and restricted cash $2m.", "Capital returns and net debt depend on unrestricted cash and leverage.", "Use unrestricted cash only for net debt; keep restricted cash separate.", "High; Q1 2026 10-Q / earnings release"),
                ("Business awards", "Business award highlights included turbo, range-extended EV, E-Powertrain and E-Cooling.", "Technology awards support the RD&E and future-content watchlist.", "Use as operating-driver context; do not treat as reported revenue.", "Medium; Q1 2026 earnings release"),
            ],
        )
        rr = _quarter_block(
            rr,
            "2025-Q4 - Quarter Notes",
            "Q4/FY2025 read: Q4 net sales $891m, Q4 adjusted EBIT $122m, Q4 adjusted FCF $139m; FY2025 net sales $3.584bn.",
            "FY2025 adjusted EBIT was $510m, adjusted FCF was $403m, buybacks were $208m and common share count declined 8% year-over-year.",
            "Watch whether initial 2026 outlook and later Q1 raise convert into quarterly adjusted EBIT and adjusted FCF.",
            "Q4 values are quarter values, not FY values; do not mix annual and quarterly non-GAAP rows.",
            [
                ("Q4 revenue", "Q4 net sales $891m; FY2025 net sales $3.584bn.", "Separates quarter actual from full-year scale.", "Use Q4 value in History_Q and FY value in guidance/promise context.", "High; Q4/FY2025 earnings release"),
                ("Q4 profit", "Q4 adjusted EBIT $122m and adjusted FCF $139m.", "Shows exit-rate profitability and cash conversion.", "Use quarter values; do not select FY columns when quarter columns exist.", "High; Q4/FY2025 earnings release"),
                ("FY2025 capital returns", "FY2025 buybacks $208m; common share count reduction 8% year-over-year.", "Capital allocation materially affected per-share metrics.", "Use as management credibility / capital-return context.", "High; Q4/FY2025 earnings release"),
                ("Initial 2026 outlook", "Initial 2026 outlook was issued with FY2025 actuals.", "Baseline for later Q1 2026 guidance raise.", "Show in Promise_Progress_UI revision log; do not force noisy OCR rows.", "High; Q4/FY2025 earnings release"),
            ],
        )
        rr = _quarter_block(
            rr,
            "2025-Q3 - Quarter Notes",
            "Q3 read: net sales $902m, adjusted EBIT $133m, adjusted EBITDA $164m, adjusted FCF $107m.",
            "Buybacks were $84m and GAAP CFO minus capex was $90m for the quarter.",
            "Watch whether product/geography mix and cash conversion support the later FY2025 exit-rate.",
            "This block is actuals-only context; no clean official guide revision is forced into the model.",
            [
                ("Revenue / mix", "Net sales $902m with product/geography cuts available in filings.", "Shows operating demand before Q4/FY2025 actuals.", "Use History_Q and Operating_Drivers; keep analytical cuts out of segment profit.", "High; 2025-Q3 filing/release package"),
                ("Adjusted operating", "Adjusted EBIT $133m and adjusted EBITDA $164m.", "Non-GAAP profitability remained a key scorecard.", "Use Valuation adjusted rows; do not replace GAAP EBITDA.", "High; Adjusted_Metrics"),
                ("Cash conversion", "Adjusted FCF $107m; CFO $100m and capex $10m.", "Cash generation supports leverage and buyback analysis.", "Show GAAP and adjusted cash-flow definitions separately.", "High; History_Q / Adjusted_Metrics"),
                ("Buybacks", "Buybacks $84m.", "Capital returns affect per-share value and cash capacity.", "Use as capital-allocation context, not operating profit.", "High; History_Q"),
            ],
        )
        rr = _quarter_block(
            rr,
            "2025-Q2 - Quarter Notes",
            "Q2 read: net sales $913m, adjusted EBIT $124m, adjusted EBITDA $154m, adjusted FCF $121m.",
            "GAAP CFO was $158m, capex was $15m and buybacks were $22m.",
            "Watch cash conversion quality and whether revenue mix supports full-year operating guidance.",
            "This block is actuals-only context; no unsupported guidance-revision row is added.",
            [
                ("Revenue / mix", "Net sales $913m.", "Sequential revenue and mix are operating context for the full-year bridge.", "Use History_Q for reported actuals and Operating_Drivers for analytical cuts.", "High; 2025-Q2 filing/release package"),
                ("Adjusted operating", "Adjusted EBIT $124m and adjusted EBITDA $154m.", "Keeps GTX's primary non-GAAP scorecard visible.", "Use Valuation adjusted rows and keep GAAP rows separate.", "High; Adjusted_Metrics"),
                ("Cash conversion", "Adjusted FCF $121m; CFO $158m and capex $15m.", "Strong cash conversion can fund debt reduction or buybacks.", "Use GAAP FCF and adjusted FCF as different definitions.", "High; History_Q / Adjusted_Metrics"),
                ("Buybacks", "Buybacks $22m.", "Quarterly capital return was lower than later quarters.", "Track in capital-return rows without changing EBITDA.", "High; History_Q"),
            ],
        )
        rr = _quarter_block(
            rr,
            "2025-Q1 - Quarter Notes",
            "Q1 read: net sales $878m, adjusted EBIT $131m, adjusted EBITDA $159m, adjusted FCF $36m.",
            "GAAP CFO was $56m, capex was $26m and buybacks were $30m.",
            "Watch early-year cash conversion before drawing conclusions from the full-year guide.",
            "This block is actuals-only context; no unsupported guidance-revision row is added.",
            [
                ("Revenue / mix", "Net sales $878m.", "Sets the first-quarter base for 2025 actuals.", "Use History_Q for reported actuals and Operating_Drivers for analytical cuts.", "High; 2025-Q1 filing/release package"),
                ("Adjusted operating", "Adjusted EBIT $131m and adjusted EBITDA $159m.", "Shows early-year non-GAAP profitability.", "Use Valuation adjusted rows and keep GAAP rows separate.", "High; Adjusted_Metrics"),
                ("Cash conversion", "Adjusted FCF $36m; CFO $56m and capex $26m.", "Cash conversion is weaker early-year context versus later quarters.", "Use GAAP FCF and adjusted FCF as different definitions.", "High; History_Q / Adjusted_Metrics"),
                ("Buybacks", "Buybacks $30m.", "Capital returns started before heavier later-year repurchases.", "Track in capital-return rows without changing EBITDA.", "High; History_Q"),
            ],
        )

        widths = {1: 20, 2: 22, 3: 26, 4: 26, 5: 26, 6: 26, 7: 26, 8: 26, 9: 26, 10: 24, 11: 24, 12: 24, 13: 22, 14: 22, 15: 46}
        for cc, width in widths.items():
            ws.column_dimensions[get_column_letter(cc)].width = width
        for row_idx in range(1, int(ws.max_row or 0) + 1):
            if ws.row_dimensions[row_idx].height is None:
                has_text = any(str(ws.cell(row_idx, cc).value or "").strip() for cc in range(1, 16))
                ws.row_dimensions[row_idx].height = 34.0 if has_text else 10.0
        ws.freeze_panes = "A8"

    if ticker_txt == "GTX":
        _fix_gtx_summary_quarter_label()
        _rewrite_gtx_promise_progress_dashboard()
        _rewrite_gtx_valuation_guidance_and_driver_panels()
        _rewrite_gtx_bs_segments_analytical_cuts()
        _clean_gtx_quarter_notes_visible_noise()

    for ws in list(wb.worksheets):
        if not _is_visible_ui_sheet(ws.title):
            continue
        max_row = int(ws.max_row or 0)
        max_col = int(ws.max_column or 0)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or val.startswith("="):
                    continue
                new_val = _normalize_cell_text(val)
                if new_val != val:
                    cell.value = new_val
        if ticker_txt == "ANF" and ws.title in {"SUMMARY", "Valuation", "BS_Segments", "Operating_Drivers", "Quarter_Notes_UI", "Promise_Progress_UI", "ANF_Investment_Case"}:
            note_text = "Quarter labels are fiscal periods; for ANF, 2025-Q4 ended 2026-01-31."
            present = False
            for rr in range(1, min(max_row, 8) + 1):
                for cc in range(1, min(max_col, 6) + 1):
                    if note_text.lower() in str(ws.cell(rr, cc).value or "").lower():
                        present = True
                        break
                if present:
                    break
            if not present and ws.title in {"Operating_Drivers", "ANF_Investment_Case", "Promise_Progress_UI", "BS_Segments"}:
                target_row = 3 if ws.title == "Operating_Drivers" else 2
                existing = str(ws.cell(target_row, 1).value or "").strip()
                if existing:
                    if "Quarter labels are fiscal periods" not in existing:
                        ws.cell(target_row, 1).value = f"{note_text} {existing}"
                else:
                    ws.cell(target_row, 1).value = note_text

        def _clamp_rows(start_row: int, min_h: float, max_h: float) -> None:
            for rr in range(start_row, max_row + 1):
                dim = ws.row_dimensions[rr]
                if bool(dim.hidden):
                    continue
                cur = dim.height
                if cur is None:
                    dim.height = min_h
                    continue
                try:
                    cur_f = float(cur)
                except Exception:
                    dim.height = min_h
                    continue
                if cur_f < min_h:
                    dim.height = min_h
                elif cur_f > max_h:
                    dim.height = max_h

        def _canonical_status_label(value: Any) -> str:
            txt = str(value or "").strip()
            low = txt.lower()
            mapping = {
                "completed": "Completed",
                "complete": "Completed",
                "delivered": "Completed",
                "achieved": "Completed",
                "hit": "Hit",
                "met": "Completed",
                "beat": "Beat",
                "miss": "Missed",
                "missed": "Missed",
                "fail": "Missed",
                "failed": "Missed",
                "met-ish": "On track",
                "met ish": "On track",
                "partial": "On track",
                "on track": "On track",
                "on_track": "On track",
                "updated": "On track",
                "raised": "On track",
                "lowered": "On track",
                "maintained": "On track",
                "open": "Open",
                "not yet measurable": "Open",
                "not yet realized": "Open",
                "mixed": "Mixed",
                "basis-dependent": "Basis-dependent",
                "basis dependent": "Basis-dependent",
                "n/a": "Not assessed",
                "na": "Not assessed",
                "not assessed": "Not assessed",
                "not assessable": "Not assessed",
            }
            return mapping.get(low, txt)

        def _status_fill_for_label(value: Any) -> Optional[PatternFill]:
            low = _canonical_status_label(value).lower()
            palette = {
                "open": "A6CEE3",
                "on track": "56B4E9",
                "completed": "009E73",
                "hit": "66C2A5",
                "beat": "66C2A5",
                "missed": "D55E00",
                "mixed": "E69F00",
                "basis-dependent": "CC79A7",
                "not assessed": "D9D9D9",
            }
            color = palette.get(low)
            return PatternFill("solid", fgColor=color) if color else None

        def _dedupe_promise_timeline_rows() -> None:
            if ws.title != "Promise_Progress_UI":
                return
            local_max_row = int(ws.max_row or 0)
            if local_max_row < 3:
                return
            required = {
                "metric",
                "previous guide",
                "new/current guide",
                "change type",
                "actual",
                "status",
            }
            active_cols: Dict[str, int] = {}
            current_block = ""
            seen: Dict[Tuple[str, ...], int] = {}
            rows_to_delete: List[int] = []
            timeline_rows: List[Dict[str, Any]] = []

            def _row_header_map(row_idx: int) -> Dict[str, int]:
                out: Dict[str, int] = {}
                for col_idx in range(1, int(ws.max_column or 0) + 1):
                    txt = str(ws.cell(row_idx, col_idx).value or "").strip().lower()
                    if txt in {"actual / latest actual", "actual / latest", "latest actual", "latest result"}:
                        txt = "actual"
                    if txt:
                        out[txt] = col_idx
                return out

            def _norm_key(v: Any) -> str:
                txt = str(v or "").strip()
                txt = re.sub(r"\s+", " ", txt)
                return txt.lower()

            for rr in range(1, local_max_row + 1):
                first_txt = str(ws.cell(rr, 1).value or "").strip()
                first_fill = str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper()
                row_map = _row_header_map(rr)
                if required.issubset(set(row_map)):
                    active_cols = row_map
                    continue
                if first_txt and first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")):
                    current_block = first_txt
                    active_cols = {}
                    continue
                if not active_cols:
                    continue
                metric_col = active_cols.get("metric")
                status_col = active_cols.get("status")
                if not metric_col or not status_col:
                    continue
                metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
                if not metric_txt:
                    continue
                # Data rows only; ignore spacer and accidental repeated headers.
                if metric_txt.lower() in required:
                    continue
                key = (
                    _norm_key(current_block),
                    _norm_key(ws.cell(rr, active_cols["metric"]).value),
                    _norm_key(ws.cell(rr, active_cols.get("horizon", 0)).value if active_cols.get("horizon") else ""),
                    _norm_key(ws.cell(rr, active_cols["previous guide"]).value),
                    _norm_key(ws.cell(rr, active_cols["new/current guide"]).value),
                    _norm_key(ws.cell(rr, active_cols["change type"]).value),
                    _norm_key(ws.cell(rr, active_cols["actual"]).value),
                    _norm_key(ws.cell(rr, active_cols.get("progress / run-rate", 0)).value if active_cols.get("progress / run-rate") else ""),
                    _norm_key(ws.cell(rr, active_cols["status"]).value),
                )
                if key in seen:
                    rows_to_delete.append(rr)
                else:
                    seen[key] = rr
                    timeline_rows.append(
                        {
                            "row": rr,
                            "block": _norm_key(current_block),
                            "metric": _norm_key(ws.cell(rr, active_cols["metric"]).value),
                            "horizon": _norm_key(
                                ws.cell(rr, active_cols.get("horizon", 0)).value
                                if active_cols.get("horizon")
                                else ""
                            ),
                            "previous": _norm_key(ws.cell(rr, active_cols["previous guide"]).value),
                            "new": _norm_key(ws.cell(rr, active_cols["new/current guide"]).value),
                            "change": _norm_key(ws.cell(rr, active_cols["change type"]).value),
                            "actual": _norm_key(ws.cell(rr, active_cols["actual"]).value),
                            "progress": _norm_key(
                                ws.cell(rr, active_cols.get("progress / run-rate", 0)).value
                                if active_cols.get("progress / run-rate")
                                else ""
                            ),
                            "status": _norm_key(ws.cell(rr, active_cols["status"]).value),
                            "source_date": _norm_key(
                                ws.cell(rr, active_cols.get("source date", 0)).value
                                if active_cols.get("source date")
                                else ""
                            ),
                            "stated_in": _norm_key(
                                ws.cell(rr, active_cols.get("stated in", 0)).value
                                if active_cols.get("stated in")
                                else ""
                            ),
                            "source_note": _norm_key(
                                ws.cell(rr, active_cols.get("source / note", 0)).value
                                if active_cols.get("source / note")
                                else ""
                            ),
                        }
                    )

            grouped: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = {}
            for item in timeline_rows:
                block = str(item.get("block") or "")
                if "revisions" not in block:
                    continue
                group_key = (
                    block,
                    str(item.get("metric") or ""),
                    str(item.get("horizon") or ""),
                )
                grouped.setdefault(group_key, []).append(item)
            def _source_date_sort_value(item: Mapping[str, Any]) -> Tuple[int, int, int]:
                try:
                    parsed = pd.Timestamp(str(item.get("source_date") or "")).date()
                    return (parsed.year, parsed.month, parsed.day)
                except Exception:
                    return (9999, 12, 31)

            def _is_real_revision(item: Mapping[str, Any]) -> bool:
                change = str(item.get("change") or "")
                prev = str(item.get("previous") or "")
                new = str(item.get("new") or "")
                if change == "initial":
                    return True
                if change in {"raised", "lowered", "narrowed"}:
                    return bool(new)
                if change == "updated":
                    return bool(prev and new and prev != new)
                return False

            def _is_carry_forward_confirmation(item: Mapping[str, Any]) -> bool:
                change = str(item.get("change") or "")
                prev = str(item.get("previous") or "")
                new = str(item.get("new") or "")
                blob = " ".join(
                    str(item.get(k) or "")
                    for k in ("actual", "source_note", "status")
                )
                if change == "maintained":
                    return True
                if change == "updated" and prev and new and prev == new:
                    return True
                if prev and new and prev == new and re.search(r"\b(confirm|confirmed|carry|carried|run-rate)\b", blob):
                    return True
                return False

            for group_rows in grouped.values():
                by_effective_guide: Dict[str, List[Dict[str, Any]]] = {}
                for item in group_rows:
                    by_effective_guide.setdefault(str(item.get("new") or ""), []).append(item)

                for same_guide_rows in by_effective_guide.values():
                    if len(same_guide_rows) <= 1:
                        item = same_guide_rows[0]
                        if (
                            _is_carry_forward_confirmation(item)
                            and re.search(r"\brun-rate\b", str(item.get("actual") or "") + " " + str(item.get("source_note") or ""))
                            and str(item.get("block") or "") != str(item.get("stated_in") or "")
                        ):
                            rows_to_delete.append(int(item["row"]))
                        continue

                    anchors = [item for item in same_guide_rows if _is_real_revision(item)]
                    if anchors:
                        for item in same_guide_rows:
                            if _is_carry_forward_confirmation(item):
                                rows_to_delete.append(int(item["row"]))
                        continue

                    carry_rows = [item for item in same_guide_rows if _is_carry_forward_confirmation(item)]
                    if len(carry_rows) > 1:
                        keep = sorted(carry_rows, key=_source_date_sort_value)[0]
                        for item in carry_rows:
                            if int(item["row"]) != int(keep["row"]):
                                rows_to_delete.append(int(item["row"]))
            rows_to_delete = sorted(set(rows_to_delete))
            for rr in reversed(rows_to_delete):
                ws.delete_rows(rr, 1)

        def _repair_promise_horizon_actuals_and_final_rows() -> None:
            if ws.title != "Promise_Progress_UI":
                return
            local_max_row = int(ws.max_row or 0)
            local_max_col = max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL)
            if local_max_row < 3:
                return
            wb_obj = getattr(ws, "parent", None)
            if wb_obj is not None and any(str(name).startswith("ANF_") for name in getattr(wb_obj, "sheetnames", [])):
                return
            first_col_values = {
                str(ws.cell(row_idx, 1).value or "").strip()
                for row_idx in range(1, min(local_max_row, 140) + 1)
            }
            if {"Tariff impact", "Real estate activity"} & first_col_values:
                return

            def _norm_header(value: Any) -> str:
                txt = str(value or "").strip().lower()
                if txt in {"actual / latest actual", "actual / latest", "latest actual", "latest result"}:
                    return "actual"
                return txt

            def _num(value: Any) -> Optional[float]:
                out = pd.to_numeric(value, errors="coerce")
                if pd.isna(out):
                    return None
                val = float(out)
                return val if math.isfinite(val) else None

            def _labels_for_history_row(qd: date, row_map: Mapping[str, Any]) -> List[str]:
                labels = {str(qd)}
                fiscal_labels: Set[str] = set()
                fiscal_label = str(row_map.get("fiscal_label") or "").strip()
                if fiscal_label:
                    fiscal_labels.add(fiscal_label)
                fy = _num(row_map.get("fiscal_year"))
                fq = _num(row_map.get("fiscal_quarter"))
                if fy is not None and fq is not None and 1 <= int(fq) <= 4:
                    fiscal_labels.add(f"{int(fy)}-Q{int(fq)}")
                if fiscal_labels:
                    labels.update(fiscal_labels)
                else:
                    labels.add(f"{qd.year}-Q{((qd.month - 1) // 3) + 1}")
                return [label for label in labels if label]

            def _year_for_history_row(qd: date, row_map: Mapping[str, Any]) -> int:
                fy = _num(row_map.get("fiscal_year"))
                return int(fy) if fy is not None else int(qd.year)

            def _build_actual_lookup() -> Tuple[Dict[str, Dict[str, float]], Dict[int, Dict[str, float]], Dict[int, date]]:
                by_period: Dict[str, Dict[str, float]] = {}
                by_year: Dict[int, Dict[str, float]] = {}
                year_end_dates: Dict[int, date] = {}
                history_period_labels_by_date: Dict[date, Set[str]] = {}

                def _add_period(labels: Iterable[str], key: str, value: Optional[float]) -> None:
                    if value is None:
                        return
                    for label in labels:
                        by_period.setdefault(label, {})[key] = value

                def _add_year(year: int, key: str, value: Optional[float]) -> None:
                    if value is None:
                        return
                    by_year.setdefault(int(year), {})[key] = by_year.setdefault(int(year), {}).get(key, 0.0) + value

                def _labels_year_or_calendar(qd: date, fiscal_year: Optional[float], fiscal_quarter: Optional[float]) -> Tuple[Set[str], int]:
                    labels: Set[str] = {str(qd)}
                    if fiscal_year is not None and fiscal_quarter is not None and 1 <= int(fiscal_quarter) <= 4:
                        labels.add(f"{int(fiscal_year)}-Q{int(fiscal_quarter)}")
                        return labels, int(fiscal_year)
                    hist_labels = history_period_labels_by_date.get(qd)
                    if hist_labels:
                        labels.update(hist_labels)
                        year_match = next((re.match(r"^(20\d{2})-Q[1-4]$", label) for label in sorted(hist_labels)), None)
                        if year_match:
                            return labels, int(year_match.group(1))
                    labels.add(f"{qd.year}-Q{((qd.month - 1) // 3) + 1}")
                    return labels, int(qd.year)

                if "History_Q" in getattr(wb, "sheetnames", []):
                    hist_ws = wb["History_Q"]
                    headers = {
                        re.sub(r"[^a-z0-9]+", "_", str(hist_ws.cell(1, cc).value or "").strip().lower()).strip("_"): cc
                        for cc in range(1, int(hist_ws.max_column or 0) + 1)
                    }
                    q_col = headers.get("quarter")
                    if q_col:
                        for row_idx in range(2, int(hist_ws.max_row or 0) + 1):
                            qd = _date_or_none(hist_ws.cell(row_idx, q_col).value)
                            if qd is None:
                                continue
                            row_map = {name: hist_ws.cell(row_idx, col_idx).value for name, col_idx in headers.items()}
                            labels = _labels_for_history_row(qd, row_map)
                            history_period_labels_by_date.setdefault(qd, set()).update(
                                label for label in labels if re.fullmatch(r"20\d{2}-Q[1-4]", str(label))
                            )
                            year = _year_for_history_row(qd, row_map)
                            year_end_dates[year] = max(year_end_dates.get(year, date.min), qd)
                            revenue = _num(row_map.get("revenue"))
                            op_income = _num(row_map.get("op_income"))
                            cfo = _num(row_map.get("cfo"))
                            capex = _num(row_map.get("capex"))
                            eps = _num(row_map.get("eps_diluted"))
                            shares = _num(row_map.get("shares_diluted"))
                            buybacks = _num(row_map.get("buybacks_cash"))
                            operating_margin = (op_income / revenue) if revenue and op_income is not None else None
                            fcf = (cfo - capex) if cfo is not None and capex is not None else None
                            for key, value in (
                                ("revenue", revenue),
                                ("operating_margin", operating_margin),
                                ("fcf", fcf),
                                ("capex", capex),
                                ("eps", eps),
                                ("shares", shares),
                                ("buybacks", buybacks),
                            ):
                                _add_period(labels, key, value)
                                _add_year(year, key, value)

                if "Adjusted_Metrics" in getattr(wb, "sheetnames", []):
                    adj_ws = wb["Adjusted_Metrics"]
                    headers = {
                        re.sub(r"[^a-z0-9]+", "_", str(adj_ws.cell(1, cc).value or "").strip().lower()).strip("_"): cc
                        for cc in range(1, int(adj_ws.max_column or 0) + 1)
                    }
                    q_col = headers.get("quarter")
                    if q_col:
                        for row_idx in range(2, int(adj_ws.max_row or 0) + 1):
                            qd = _date_or_none(adj_ws.cell(row_idx, q_col).value)
                            if qd is None:
                                continue
                            period_type = str(adj_ws.cell(row_idx, headers.get("period_type", 0)).value or "").strip().lower() if headers.get("period_type") else ""
                            if period_type and "annual" in period_type:
                                continue
                            row_map = {name: adj_ws.cell(row_idx, col_idx).value for name, col_idx in headers.items()}
                            fiscal_year = _num(row_map.get("fiscal_year"))
                            fiscal_quarter = _num(row_map.get("fiscal_quarter"))
                            labels, year = _labels_year_or_calendar(qd, fiscal_year, fiscal_quarter)
                            year_end_dates[year] = max(year_end_dates.get(year, date.min), qd)
                            for key, source_name in (
                                ("adj_ebit", "adj_ebit"),
                                ("adj_ebitda", "adj_ebitda"),
                                ("adj_eps", "adj_eps"),
                                ("adj_fcf", "adj_fcf"),
                            ):
                                value = _num(row_map.get(source_name))
                                _add_period(labels, key, value)
                                _add_year(year, key, value)
                if "Valuation" in getattr(wb, "sheetnames", []):
                    val_ws = wb["Valuation"]
                    qcols: Dict[str, int] = {}
                    for col_idx in range(1, int(val_ws.max_column or 0) + 1):
                        label = str(val_ws.cell(6, col_idx).value or "").strip()
                        if re.fullmatch(r"20\d{2}-Q[1-4]", label):
                            qcols[label] = col_idx
                    val_row_by_label = {
                        str(val_ws.cell(row_idx, 1).value or "").strip(): row_idx
                        for row_idx in range(1, int(val_ws.max_row or 0) + 1)
                        if str(val_ws.cell(row_idx, 1).value or "").strip()
                    }
                    for row_label, actual_key, scale in (
                        ("Adj EPS", "adj_eps", 1.0),
                        ("Adj EBITDA", "adj_ebitda", 1_000_000.0),
                    ):
                        row_idx = val_row_by_label.get(row_label)
                        if not row_idx:
                            continue
                        for label, col_idx in qcols.items():
                            vv = _num(val_ws.cell(row_idx, col_idx).value)
                            if vv is None:
                                continue
                            _add_period([label], actual_key, float(vv) * float(scale))
                    for row_label, actual_key, scale in (
                        ("Adj EBIT (TTM)", "adj_ebit", 1_000_000.0),
                        ("Adj EBITDA (TTM)", "adj_ebitda", 1_000_000.0),
                        ("Adj EPS (TTM)", "adj_eps", 1.0),
                    ):
                        row_idx = val_row_by_label.get(row_label)
                        if not row_idx:
                            continue
                        for label, col_idx in qcols.items():
                            year_match = re.match(r"^(20\d{2})-Q4$", label)
                            if not year_match:
                                continue
                            vv = _num(val_ws.cell(row_idx, col_idx).value)
                            if vv is None:
                                continue
                            by_year.setdefault(int(year_match.group(1)), {})[actual_key] = float(vv) * float(scale)
                return by_period, by_year, year_end_dates

            actuals_by_period, actuals_by_year, year_end_dates = _build_actual_lookup()

            def _metric_actual_key(metric_in: Any) -> str:
                metric_low = str(metric_in or "").strip().lower()
                if "adjusted ebitda" in metric_low or "adj ebitda" in metric_low:
                    return "adj_ebitda"
                if "adjusted ebit" in metric_low or "adj ebit" in metric_low:
                    return "adj_ebit"
                if "fcf" in metric_low or "free cash" in metric_low:
                    return "adj_fcf" if "adjusted" in metric_low or "adj" in metric_low else "fcf"
                if "capex" in metric_low or "capital expenditure" in metric_low:
                    return "capex"
                if "operating margin" in metric_low or "op margin" in metric_low:
                    return "operating_margin"
                if "eps" in metric_low:
                    return "adj_eps" if "adjusted" in metric_low or "adj" in metric_low else "eps"
                if "share" in metric_low and "repurchase" not in metric_low and "buyback" not in metric_low:
                    return "shares"
                if "repurchase" in metric_low or "buyback" in metric_low:
                    return "buybacks"
                if "revenue" in metric_low or "sales" in metric_low:
                    return "revenue"
                return ""

            def _format_actual(metric_in: Any, key: str, value: Any) -> str:
                val = pd.to_numeric(value, errors="coerce")
                if pd.isna(val):
                    return ""
                num = float(val)
                if key == "operating_margin":
                    return f"{num * 100:.1f}%"
                if key in {"adj_eps", "eps"}:
                    return f"${num:,.2f}"
                if key == "shares":
                    return f"{num / 1_000_000:,.1f}m" if abs(num) > 1_000_000 else f"{num:,.1f}m"
                abs_num = abs(num)
                if abs_num >= 1_000_000_000:
                    return f"${num / 1_000_000_000:,.2f}bn"
                if abs_num >= 1_000_000:
                    return f"${num / 1_000_000:,.1f}m"
                return f"${num:,.1f}"

            def _quarter_score(label: Any) -> Tuple[Optional[int], Optional[int]]:
                m = re.search(r"\b(20\d{2})-Q([1-4])\b", str(label or ""), flags=re.I)
                if not m:
                    return None, None
                return int(m.group(1)), int(m.group(2))

            def _annual_year(label: Any) -> Optional[int]:
                m = re.fullmatch(r"(20\d{2})\s+year", str(label or "").strip(), flags=re.I)
                return int(m.group(1)) if m else None

            def _actual_for_period(metric: Any, period_label: Any) -> str:
                key = _metric_actual_key(metric)
                if not key:
                    return ""
                period_vals = actuals_by_period.get(str(period_label or "").strip(), {})
                if key in period_vals:
                    return _format_actual(metric, key, period_vals[key])
                if key == "eps" and "adj_eps" in period_vals:
                    return _format_actual(metric, "adj_eps", period_vals["adj_eps"])
                if key == "adj_fcf" and "fcf" in period_vals:
                    return _format_actual(metric, "fcf", period_vals["fcf"])
                return ""

            def _actual_for_year(metric: Any, year: int) -> str:
                key = _metric_actual_key(metric)
                if not key:
                    return ""
                annual_vals = actuals_by_year.get(int(year), {})
                if key in annual_vals:
                    return _format_actual(metric, key, annual_vals[key])
                if key == "eps" and "adj_eps" in annual_vals:
                    return _format_actual(metric, "adj_eps", annual_vals["adj_eps"])
                if key == "adj_fcf" and "fcf" in annual_vals:
                    return _format_actual(metric, "fcf", annual_vals["fcf"])
                return ""

            def _ytd_for_year_to_quarter(metric: Any, year: int, quarter_num: int) -> str:
                key = _metric_actual_key(metric)
                if not key or quarter_num not in {1, 2, 3, 4}:
                    return ""
                if key in {"operating_margin", "shares"}:
                    return ""
                vals: List[float] = []
                actual_key = key
                for idx in range(1, int(quarter_num) + 1):
                    period_vals = actuals_by_period.get(f"{int(year)}-Q{idx}", {})
                    if actual_key not in period_vals and actual_key == "eps" and "adj_eps" in period_vals:
                        actual_key = "adj_eps"
                    if actual_key not in period_vals and actual_key == "adj_fcf" and "fcf" in period_vals:
                        actual_key = "fcf"
                    if actual_key not in period_vals:
                        return ""
                    vv = _num(period_vals.get(actual_key))
                    if vv is None:
                        return ""
                    vals.append(float(vv))
                return _format_actual(metric, actual_key, sum(vals))

            def _numbers(txt: Any) -> List[float]:
                vals: List[float] = []
                for match in re.findall(r"(?<![A-Za-z])\d+(?:,\d{3})*(?:\.\d+)?", str(txt or "")):
                    try:
                        vals.append(float(match.replace(",", "")))
                    except Exception:
                        continue
                return vals

            def _status_from_guidance_actual(metric: Any, guide: Any, actual: Any) -> str:
                metric_low = str(metric or "").lower()
                guide_nums = _numbers(guide)
                actual_nums = _numbers(actual)
                if not actual_nums:
                    return ""
                if not guide_nums:
                    return "Completed"
                actual_val = actual_nums[0]
                low = min(guide_nums)
                high = max(guide_nums)
                if len(guide_nums) == 1:
                    low = high = guide_nums[0]
                tolerance = max(abs(high) * 0.001, 0.001)
                if "capex" in metric_low or "capital expenditure" in metric_low:
                    return "Hit" if low - tolerance <= actual_val <= high + tolerance else "Mixed"
                if "share" in metric_low and "repurchase" not in metric_low and "buyback" not in metric_low:
                    return "Hit" if low - tolerance <= actual_val <= high + tolerance else "Mixed"
                if actual_val > high + tolerance:
                    return "Beat"
                if actual_val < low - tolerance:
                    return "Missed"
                return "Hit"

            def _run_rate_actual_from_text(txt: Any) -> str:
                match = re.search(r"\$?\s*(\d+(?:\.\d+)?)\s*m\s+run[- ]rate", str(txt or ""), flags=re.I)
                if not match:
                    return ""
                return f"${float(match.group(1)):g}m run-rate"

            active_cols: Dict[str, int] = {}
            current_block = ""
            rows_to_delete: List[int] = []
            annual_latest_by_metric: Dict[Tuple[int, str], Dict[str, Any]] = {}
            annual_q4_rows: Set[Tuple[int, str]] = set()
            q4_blocks: Dict[int, Dict[str, Any]] = {}
            for row_idx in range(1, local_max_row + 1):
                first_txt = str(ws.cell(row_idx, 1).value or "").strip()
                first_fill = str(ws.cell(row_idx, 1).fill.fgColor.rgb or "").upper()
                row_map = {
                    _norm_header(ws.cell(row_idx, col_idx).value): col_idx
                    for col_idx in range(1, local_max_col + 1)
                    if str(ws.cell(row_idx, col_idx).value or "").strip()
                }
                if "actual" in row_map and ("metric" in row_map or "milestone" in row_map):
                    active_cols = row_map
                    continue
                if first_txt and first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")):
                    current_block = first_txt
                    active_cols = {}
                    continue
                if not active_cols:
                    continue
                metric_col = active_cols.get("metric") or active_cols.get("milestone")
                actual_col = active_cols.get("actual")
                progress_col = active_cols.get("progress / run-rate")
                status_col = active_cols.get("status")
                note_col = active_cols.get("source / note") or active_cols.get("notes/source")
                new_col = active_cols.get("new/current guide")
                if not metric_col or not actual_col:
                    continue
                metric_txt = str(ws.cell(row_idx, metric_col).value or "").strip()
                if not metric_txt or metric_txt.lower() in {"metric", "milestone"}:
                    continue
                useful_values = [
                    ws.cell(row_idx, col_idx).value
                    for col_idx in range(1, min(local_max_col, PROMISE_VISIBLE_MAX_COL) + 1)
                    if col_idx != metric_col
                ]
                if all(str(value or "").strip() == "" for value in useful_values):
                    rows_to_delete.append(row_idx)
                    continue

                note_txt = str(ws.cell(row_idx, note_col).value or "").strip() if note_col else ""
                actual_txt = str(ws.cell(row_idx, actual_col).value or "").strip()
                new_txt = str(ws.cell(row_idx, new_col).value or "").strip() if new_col else ""
                if metric_txt.lower() == "debt reduction" and status_col:
                    if ticker_txt == "GTX":
                        row_blob = f"{new_txt} {actual_txt} {note_txt}"
                        if (
                            re.search(r"\$?\s*50(?:\.0)?\s*m\b", row_blob, flags=re.I)
                            and re.search(r"\b(disclosed|repay|repayment|repricing|term\s+loan|may\s+18)\b", row_blob, flags=re.I)
                        ):
                            if new_col:
                                ws.cell(row_idx, new_col).value = "Post-quarter event-context / pro-forma: $50m term-loan repayment/repricing"
                            ws.cell(row_idx, actual_col).value = "$50m disclosed; post-quarter/pro-forma, not Q1 reported history"
                            if note_col:
                                ws.cell(row_idx, note_col).value = (
                                    "Post-quarter May 18 2026 term-loan repayment/repricing; "
                                    "valuation/debt-event context only, not reported Q1 history."
                                )
                    status_txt = _canonical_status_label(ws.cell(row_idx, status_col).value)
                    if not actual_txt and status_txt == "Completed" and re.search(r"\b(repay|repaid|paid off|pay down|paydown)\b", note_txt, flags=re.I):
                        ws.cell(row_idx, actual_col).value = "Debt repaid"
                    continue

                if metric_txt.lower() == "cost savings target":
                    stated_col = active_cols.get("stated in")
                    source_col = active_cols.get("source date")
                    stated_txt = str(ws.cell(row_idx, stated_col).value or current_block).strip() if stated_col else current_block
                    source_txt = str(ws.cell(row_idx, source_col).value or "").strip() if source_col else ""
                    stated_year, _ = _quarter_score(stated_txt)
                    source_year = None
                    try:
                        source_year = pd.Timestamp(source_txt).year if source_txt else None
                    except Exception:
                        source_year = None
                    if (stated_year and stated_year >= 2025) or (source_year and source_year >= 2025):
                        run_rate = _run_rate_actual_from_text(" ".join([actual_txt, note_txt]))
                        if run_rate:
                            ws.cell(row_idx, actual_col).value = ""
                            if progress_col:
                                ws.cell(row_idx, progress_col).value = _promise_progress_label(run_rate, metric=metric_txt, stated=stated_txt)
                    continue

                horizon_col = active_cols.get("horizon")
                stated_col = active_cols.get("stated in")
                new_col = active_cols.get("new/current guide")
                prev_col = active_cols.get("previous guide")
                source_col = active_cols.get("source date")
                if not (horizon_col and stated_col and new_col and status_col):
                    continue

                horizon_txt = str(ws.cell(row_idx, horizon_col).value or "").strip()
                stated_txt = str(ws.cell(row_idx, stated_col).value or "").strip()
                source_txt = str(ws.cell(row_idx, source_col).value or "").strip() if source_col else ""
                year = _annual_year(horizon_txt)
                stated_year, stated_q = _quarter_score(stated_txt)
                if year is not None and stated_year is not None:
                    if stated_year == year and stated_q in {1, 2, 3}:
                        quarter_actual = _actual_for_period(metric_txt, stated_txt)
                        if quarter_actual:
                            ws.cell(row_idx, actual_col).value = quarter_actual
                            if progress_col:
                                ytd_actual = _ytd_for_year_to_quarter(metric_txt, year, stated_q)
                                ws.cell(row_idx, progress_col).value = f"YTD: {ytd_actual}" if ytd_actual else _promise_progress_label(quarter_actual, metric=metric_txt, stated=stated_txt)
                            ws.cell(row_idx, status_col).value = "On track"
                        elif actual_txt and re.search(r"\bbn|\$|m\b|\d", actual_txt, flags=re.I):
                            ws.cell(row_idx, actual_col).value = ""
                            ws.cell(row_idx, status_col).value = "Open"
                    elif stated_year == year and stated_q == 4:
                        annual_actual = _actual_for_year(metric_txt, year)
                        actual_key_for_row = _metric_actual_key(metric_txt)
                        force_annual_actual = (
                            actual_key_for_row in {"revenue", "adj_ebit", "adj_ebitda", "fcf", "capex", "buybacks"}
                            and not re.search(r"\b(growth|margin|rate|bps|basis|share count|diluted shares)\b", metric_txt, flags=re.I)
                        )
                        if annual_actual and (not actual_txt or force_annual_actual):
                            ws.cell(row_idx, actual_col).value = annual_actual
                            guide = str(ws.cell(row_idx, new_col).value or ws.cell(row_idx, prev_col).value or "").strip() if prev_col else str(ws.cell(row_idx, new_col).value or "").strip()
                            status = _status_from_guidance_actual(metric_txt, guide, annual_actual)
                            if status:
                                ws.cell(row_idx, status_col).value = status
                        annual_q4_rows.add((year, metric_txt.lower()))
                    elif stated_year < year:
                        if actual_txt and re.search(r"\bbn|\$|m\b|\d", actual_txt, flags=re.I):
                            ws.cell(row_idx, actual_col).value = ""
                            ws.cell(row_idx, status_col).value = "Open"
                    if stated_year is not None and stated_year < year:
                        key = (year, metric_txt.lower())
                        source_dt = date.min
                        try:
                            source_dt = pd.Timestamp(source_txt).date()
                        except Exception:
                            pass
                        prior = annual_latest_by_metric.get(key)
                        if prior is None or source_dt >= prior.get("source_dt", date.min):
                            annual_latest_by_metric[key] = {
                                "metric": metric_txt,
                                "guide": str(ws.cell(row_idx, new_col).value or ws.cell(row_idx, prev_col).value or "").strip() if prev_col else str(ws.cell(row_idx, new_col).value or "").strip(),
                                "source_dt": source_dt,
                            }
                    elif stated_year == year and stated_q in {1, 2, 3}:
                        key = (year, metric_txt.lower())
                        source_dt = date.min
                        try:
                            source_dt = pd.Timestamp(source_txt).date()
                        except Exception:
                            pass
                        prior = annual_latest_by_metric.get(key)
                        if prior is None or source_dt >= prior.get("source_dt", date.min):
                            annual_latest_by_metric[key] = {
                                "metric": metric_txt,
                                "guide": str(ws.cell(row_idx, new_col).value or ws.cell(row_idx, prev_col).value or "").strip() if prev_col else str(ws.cell(row_idx, new_col).value or "").strip(),
                                "source_dt": source_dt,
                            }

                if str(current_block).endswith("revisions"):
                    block_year, block_q = _quarter_score(current_block)
                    if block_year and block_q == 4:
                        q4_blocks[block_year] = {
                            "start": q4_blocks.get(block_year, {}).get("start", row_idx),
                            "end": row_idx,
                            "source_date": source_txt or str(year_end_dates.get(block_year) or date(block_year, 12, 31)),
                        }

            for row_idx in sorted(set(rows_to_delete), reverse=True):
                ws.delete_rows(row_idx, 1)

        def _clean_promise_timeline_lifecycle_values() -> None:
            if ws.title != "Promise_Progress_UI":
                return
            local_max_row = int(ws.max_row or 0)
            local_max_col = max(int(ws.max_column or 0), PROMISE_VISIBLE_MAX_COL)
            if local_max_row < 3:
                return
            current_block = ""
            active_cols: Dict[str, int] = {}

            def _row_header_map(row_idx: int) -> Dict[str, int]:
                out: Dict[str, int] = {}
                for col_idx in range(1, local_max_col + 1):
                    txt = str(ws.cell(row_idx, col_idx).value or "").strip().lower()
                    if txt in {"actual / latest actual", "actual / latest", "latest actual", "latest result"}:
                        txt = "actual"
                    if txt:
                        out[txt] = col_idx
                return out

            for rr in range(1, local_max_row + 1):
                first_txt = str(ws.cell(rr, 1).value or "").strip()
                first_fill = str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper()
                row_map = _row_header_map(rr)
                if {"metric", "new/current guide", "change type"}.issubset(set(row_map)):
                    active_cols = row_map
                    continue
                if first_txt and first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")):
                    current_block = first_txt
                    active_cols = {}
                    continue
                if not active_cols or "revisions" not in current_block.lower():
                    continue
                metric_col = active_cols.get("metric")
                if not metric_col:
                    continue
                metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
                if "cost savings target" not in metric_txt.lower():
                    continue
                actual_col = active_cols.get("actual")
                progress_col = active_cols.get("progress / run-rate")
                note_col = active_cols.get("source / note")
                change_col = active_cols.get("change type")
                prev_col = active_cols.get("previous guide")
                new_col = active_cols.get("new/current guide")
                stated_col = active_cols.get("stated in")
                if not (actual_col and note_col and change_col and new_col):
                    continue
                actual_txt = str(ws.cell(rr, actual_col).value or "")
                note_txt = str(ws.cell(rr, note_col).value or "")
                blob = f"{actual_txt} {note_txt}"
                if not re.search(r"\b(run-rate|latest disclosed|latest run|2026-03-31)\b", blob, flags=re.I):
                    continue
                prev_txt = str(ws.cell(rr, prev_col).value or "").strip() if prev_col else ""
                new_txt = str(ws.cell(rr, new_col).value or "").strip()
                change_txt = _shared_visible_period_text(str(ws.cell(rr, change_col).value or "")).strip()
                stated_txt = str(ws.cell(rr, stated_col).value or "").strip() if stated_col else ""
                if not stated_txt:
                    stated_txt = re.sub(r"\s+revisions\s*$", "", current_block, flags=re.I).strip()
                stated_year_match = re.search(r"\b(20\d{2})-Q[1-4]\b", stated_txt)
                stated_year = int(stated_year_match.group(1)) if stated_year_match else 0
                run_rate_match = re.search(r"\$?\s*(\d+(?:\.\d+)?)\s*m\s+run[- ]rate", blob, flags=re.I)
                if run_rate_match and stated_year >= 2025:
                    run_rate_txt = f"${float(run_rate_match.group(1)):g}m run-rate"
                    ws.cell(rr, actual_col).value = ""
                    if progress_col:
                        ws.cell(rr, progress_col).value = _promise_progress_label(run_rate_txt, metric=metric_txt, stated=stated_txt)
                    continue
                ws.cell(rr, actual_col).value = ""
                if progress_col and not str(ws.cell(rr, progress_col).value or "").strip():
                    ws.cell(rr, progress_col).value = "not yet measurable"
                change_low = change_txt.lower()
                if change_low == "initial" or not prev_txt:
                    clean_note = f"Initial {stated_txt} cost savings target."
                elif prev_txt and new_txt and prev_txt != new_txt:
                    verb = {
                        "raised": "Raised",
                        "lowered": "Lowered",
                        "narrowed": "Narrowed",
                        "updated": "Updated",
                    }.get(change_low, "Updated")
                    clean_note = f"{verb} cost savings target to {new_txt}."
                elif change_low == "maintained" and new_txt:
                    clean_note = f"Maintained cost savings target at {new_txt}."
                else:
                    clean_note = f"Cost savings target update: {new_txt}."
                ws.cell(rr, note_col).value = clean_note

        def _evaluate_completed_promise_timeline_statuses() -> None:
            if ws.title != "Promise_Progress_UI":
                return
            local_max_row = int(ws.max_row or 0)
            local_max_col = int(ws.max_column or 0)
            if local_max_row < 3:
                return
            current_block = ""
            active_cols: Dict[str, int] = {}

            def _row_header_map(row_idx: int) -> Dict[str, int]:
                out: Dict[str, int] = {}
                for col_idx in range(1, local_max_col + 1):
                    txt = str(ws.cell(row_idx, col_idx).value or "").strip().lower()
                    if txt in {"actual / latest actual", "actual / latest", "latest actual", "latest result"}:
                        txt = "actual"
                    if txt:
                        out[txt] = col_idx
                return out

            def _numbers(txt: Any) -> List[float]:
                raw = str(txt or "")
                vals: List[float] = []
                for match in re.findall(r"(?<![A-Za-z])\d+(?:,\d{3})*(?:\.\d+)?", raw):
                    try:
                        vals.append(float(match.replace(",", "")))
                    except Exception:
                        continue
                return vals

            def _horizon_is_complete(txt: Any) -> bool:
                horizon = str(txt or "")
                m = re.search(r"\b(20\d{2})\s+year\b", horizon, flags=re.I)
                if m:
                    return int(m.group(1)) <= 2025
                m = re.search(r"\b(20\d{2})-Q([1-4])\b", horizon, flags=re.I)
                if m:
                    year = int(m.group(1))
                    quarter = int(m.group(2))
                    return (year, quarter) <= (2025, 4)
                return False

            def _status_from_actual(metric: Any, guide: Any, actual: Any, note: Any) -> str:
                metric_low = str(metric or "").lower()
                actual_txt = str(actual or "")
                note_txt = str(note or "")
                basis_blob = f"{actual_txt} {note_txt}"
                basis_ambiguous = bool(
                    re.search(r"\b(basis differs|basis-dependent|gaap\s*/|adjusted\s*/|gaap.*adjusted|adjusted.*gaap)\b", basis_blob, flags=re.I)
                )
                if basis_ambiguous and not ("adjusted eps" in metric_low and re.search(r"\badjusted\b", actual_txt, flags=re.I)):
                    return "Basis-dependent"
                guide_nums = _numbers(guide)
                actual_nums = _numbers(actual_txt)
                if not actual_nums:
                    return "Not assessed"
                actual_val = actual_nums[0]
                if not guide_nums:
                    return "Hit"
                low = min(guide_nums)
                high = max(guide_nums)
                if len(guide_nums) == 1:
                    low = high = guide_nums[0]
                tolerance = max(abs(high) * 0.03, 0.1)
                if "capex" in metric_low or "capital expenditure" in metric_low:
                    return "Hit" if low - tolerance <= actual_val <= high + tolerance else "Mixed"
                if "share" in metric_low and "repurchase" not in metric_low:
                    return "Hit" if low - tolerance <= actual_val <= high + tolerance else "Mixed"
                if actual_val > high + tolerance:
                    return "Beat"
                if actual_val < low - tolerance:
                    return "Missed"
                return "Hit"

            for rr in range(1, local_max_row + 1):
                first_txt = str(ws.cell(rr, 1).value or "").strip()
                first_fill = str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper()
                row_map = _row_header_map(rr)
                if {"metric", "new/current guide", "actual", "status", "horizon"}.issubset(set(row_map)):
                    active_cols = row_map
                    continue
                if first_txt and first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")):
                    current_block = first_txt
                    active_cols = {}
                    continue
                if not active_cols or "revisions" not in current_block.lower():
                    continue
                metric_col = active_cols.get("metric")
                new_col = active_cols.get("new/current guide")
                actual_col = active_cols.get("actual")
                status_col = active_cols.get("status")
                horizon_col = active_cols.get("horizon")
                note_col = active_cols.get("source / note")
                if not (metric_col and new_col and actual_col and status_col and horizon_col):
                    continue
                metric_txt = str(ws.cell(rr, metric_col).value or "").strip()
                if not metric_txt or metric_txt.lower() == "metric":
                    continue
                if "facility qualification" in metric_txt.lower():
                    continue
                status_txt = _canonical_status_label(ws.cell(rr, status_col).value)
                if status_txt not in {"Open", "On track"}:
                    continue
                actual_txt = str(ws.cell(rr, actual_col).value or "").strip()
                if not actual_txt or re.search(r"\b(not yet|not measurable|not assessed|open|expected|expected in|expected to)\b", actual_txt, flags=re.I):
                    continue
                if "cost savings target" in metric_txt.lower() and re.search(r"\brun[- ]rate\b", actual_txt, flags=re.I):
                    ws.cell(rr, status_col).value = "On track"
                    continue
                horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip()
                stated_txt = str(ws.cell(rr, active_cols.get("stated in", 0)).value or "").strip() if active_cols.get("stated in") else ""
                annual_match = re.fullmatch(r"(20\d{2})\s+year", horizon_txt, flags=re.I)
                stated_match = re.search(r"\b(20\d{2})-Q([1-4])\b", stated_txt, flags=re.I)
                if annual_match and stated_match and int(annual_match.group(1)) == int(stated_match.group(1)) and int(stated_match.group(2)) in {1, 2, 3}:
                    continue
                if not _horizon_is_complete(horizon_txt):
                    continue
                new_txt = str(ws.cell(rr, new_col).value or "").strip()
                note_txt = str(ws.cell(rr, note_col).value or "").strip() if note_col else ""
                ws.cell(rr, status_col).value = _status_from_actual(metric_txt, new_txt, actual_txt, note_txt)

        def _repair_promise_headers_and_remove_blank_rows() -> None:
            if ws.title != "Promise_Progress_UI":
                return
            local_max_row = int(ws.max_row or 0)
            local_max_col = int(ws.max_column or 0)
            current_block = ""
            active_cols: Dict[str, int] = {}
            rows_to_delete: List[int] = []
            timeline_headers = PROMISE_TIMELINE_HEADERS
            annual_headers = [
                "Metric",
                "Initial guide",
                "Q1 update",
                "Q2 update",
                "Q3 update",
                "Q4 update",
                "Actual",
                "Status",
                "Notes/source",
                "",
                "",
                "",
            ]
            open_headers = ["Metric", "Current guide", "Horizon", "Status", "Notes/source", "", "", "", "", "", "", ""]
            milestone_headers = ["Milestone", "Target / plan", "Actual", "Status", "Notes/source", "", "", "", "", "", "", ""]

            def _header_map(row_idx: int) -> Dict[str, int]:
                out: Dict[str, int] = {}
                for col_idx in range(1, local_max_col + 1):
                    txt = str(ws.cell(row_idx, col_idx).value or "").strip().lower()
                    if txt in {"actual / latest actual", "actual / latest", "latest actual", "latest result"}:
                        txt = "actual"
                    if txt:
                        out[txt] = col_idx
                return out

            def _unmerge_row(row_idx: int) -> None:
                row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, min(local_max_col, PROMISE_VISIBLE_MAX_COL) + 1)]
                for merge_range in list(ws.merged_cells.ranges):
                    if merge_range.min_row == row_idx and merge_range.max_row == row_idx:
                        ws.unmerge_cells(str(merge_range))
                for col_idx, value in enumerate(row_values, start=1):
                    if value not in (None, ""):
                        ws.cell(row_idx, col_idx).value = value

            for row_idx in range(1, local_max_row + 1):
                first_txt = str(ws.cell(row_idx, 1).value or "").strip()
                first_fill = str(ws.cell(row_idx, 1).fill.fgColor.rgb or "").upper()
                if first_txt and first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")):
                    current_block = first_txt
                    active_cols = {}
                    continue
                if first_txt in {"Metric", "Milestone"}:
                    if current_block.endswith("revisions"):
                        labels = timeline_headers
                    elif current_block.endswith("milestone progression"):
                        labels = milestone_headers
                    elif current_block.endswith("guidance progression"):
                        labels = annual_headers
                    elif current_block.endswith("open guidance") or current_block == "Open guidance":
                        labels = open_headers
                    else:
                        labels = []
                    if labels:
                        _unmerge_row(row_idx)
                        for col_idx, label in enumerate(labels, start=1):
                            ws.cell(row_idx, col_idx).value = label
                    active_cols = _header_map(row_idx)
                    continue
                if not active_cols:
                    continue
                metric_col = active_cols.get("metric") or active_cols.get("milestone")
                if not metric_col:
                    continue
                metric_txt = str(ws.cell(row_idx, metric_col).value or "").strip()
                if not metric_txt or metric_txt.lower() in {"metric", "milestone"}:
                    continue
                other_values = [
                    ws.cell(row_idx, col_idx).value
                    for col_idx in range(1, min(local_max_col, PROMISE_VISIBLE_MAX_COL) + 1)
                    if col_idx != metric_col
                ]
                if all(str(value or "").strip() == "" for value in other_values):
                    rows_to_delete.append(row_idx)
            for row_idx in sorted(set(rows_to_delete), reverse=True):
                ws.delete_rows(row_idx, 1)

        def _standardize_promise_status_cells() -> None:
            if ws.title != "Promise_Progress_UI":
                return
            _repair_promise_horizon_actuals_and_final_rows()
            _dedupe_promise_timeline_rows()
            _clean_promise_timeline_lifecycle_values()
            _evaluate_completed_promise_timeline_statuses()
            _repair_promise_headers_and_remove_blank_rows()
            exact_status_values = {
                "completed",
                "complete",
                "delivered",
                "achieved",
                "met",
                "hit",
                "beat",
                "miss",
                "missed",
                "fail",
                "failed",
                "met-ish",
                "met ish",
                "partial",
                "on track",
                "on_track",
                "open",
                "not yet measurable",
                "not yet realized",
                "mixed",
                "basis-dependent",
                "basis dependent",
                "n/a",
                "na",
                "not assessed",
                "not assessable",
            }
            status_cols: Set[int] = set()
            change_cols: Set[int] = set()
            actual_cols: Set[int] = set()
            for rr in range(1, max_row + 1):
                for cc in range(1, max_col + 1):
                    header = str(ws.cell(rr, cc).value or "").strip().lower()
                    if header in {"status", "result", "actual/status"}:
                        status_cols.add(cc)
                    elif header == "change type":
                        change_cols.add(cc)
                    elif header in {"actual / latest actual", "actual", "latest actual", "latest result"}:
                        actual_cols.add(cc)
            if not status_cols:
                return

            neutral_fill = PatternFill("solid", fgColor="FFFFFF")
            neutral_alt_fill = PatternFill("solid", fgColor="F6F9FC")

            def _is_section_or_header_row(row_idx: int) -> bool:
                first_fill = str(ws.cell(row_idx, 1).fill.fgColor.rgb or "").upper()
                row_vals = {
                    str(ws.cell(row_idx, cc).value or "").strip().lower()
                    for cc in range(1, min(max_col, 10) + 1)
                }
                return (
                    first_fill.endswith(("5B9BD5", "6FA8DC", "4472C4"))
                    or ("metric" in row_vals and ("status" in row_vals or "change type" in row_vals))
                )

            header_roles: List[Tuple[int, Dict[int, str]]] = []
            for hdr_row in range(1, max_row + 1):
                roles: Dict[int, str] = {}
                row_vals = {
                    str(ws.cell(hdr_row, cc).value or "").strip().lower()
                    for cc in range(1, min(max_col, 10) + 1)
                }
                if "metric" not in row_vals:
                    continue
                for cc in range(1, max_col + 1):
                    header = str(ws.cell(hdr_row, cc).value or "").strip().lower()
                    if header in {"status", "result", "actual/status"}:
                        roles[cc] = "status"
                    elif header == "change type":
                        roles[cc] = "change"
                    elif header in {"actual / latest actual", "actual", "latest actual", "latest result"}:
                        roles[cc] = "actual"
                if roles:
                    header_roles.append((hdr_row, roles))

            def _role_for_cell(row_idx: int, col_idx: int) -> str:
                role = ""
                for hdr_row, roles in header_roles:
                    if hdr_row > row_idx:
                        break
                    role = roles.get(col_idx, "")
                return role

            def _neutralize_non_status_column(row_idx: int, col_idx: int) -> None:
                if _is_section_or_header_row(row_idx):
                    return
                ws.cell(row_idx, col_idx).fill = neutral_alt_fill if row_idx % 2 else neutral_fill

            change_map = {
                "initial": "Initial",
                "updated": "Updated",
                "raised": "Raised",
                "lowered": "Lowered",
                "narrowed": "Narrowed",
                "maintained": "Maintained",
            }
            for rr in range(1, max_row + 1):
                if _is_section_or_header_row(rr):
                    continue
                for cc in range(1, max_col + 1):
                    role = _role_for_cell(rr, cc)
                    if not role:
                        continue
                    cell = ws.cell(rr, cc)
                    raw = str(cell.value or "").strip()
                    low = raw.lower()
                    if role == "change":
                        if low in change_map:
                            cell.value = change_map[low]
                        elif "raised" in low:
                            cell.value = "Raised"
                        elif "lowered" in low:
                            cell.value = "Lowered"
                        elif "narrow" in low:
                            cell.value = "Narrowed"
                        elif "maintain" in low:
                            cell.value = "Maintained"
                        elif low in exact_status_values:
                            cell.value = "Initial" if low == "open" else "Updated"
                        _neutralize_non_status_column(rr, cc)
                    elif role == "actual":
                        _neutralize_non_status_column(rr, cc)
                    elif role == "status":
                        canonical = _canonical_status_label(raw)
                        if canonical and canonical != raw:
                            cell.value = canonical
                        fill = _status_fill_for_label(canonical)
                        if fill is not None:
                            cell.fill = fill
            for rr in range(1, max_row + 1):
                for cc in range(1, max_col + 1):
                    if _role_for_cell(rr, cc) in {"status", "change", "actual"}:
                        continue
                    cell = ws.cell(rr, cc)
                    raw = str(cell.value or "").strip()
                    if raw.lower() not in exact_status_values:
                        continue
                    canonical = _canonical_status_label(raw)
                    if canonical and canonical != raw:
                        cell.value = canonical

            active_cols: Dict[str, int] = {}
            status_like_progress = {
                "open",
                "on track",
                "completed",
                "complete",
                "hit",
                "beat",
                "missed",
                "miss",
                "mixed",
                "basis-dependent",
                "basis dependent",
                "not assessed",
            }
            for rr in range(1, max_row + 1):
                row_map = {
                    str(ws.cell(rr, cc).value or "").strip().lower(): cc
                    for cc in range(1, max_col + 1)
                    if str(ws.cell(rr, cc).value or "").strip()
                }
                if {"metric", "progress / run-rate", "status"}.issubset(set(row_map)):
                    active_cols = row_map
                    continue
                if _is_section_or_header_row(rr):
                    active_cols = {}
                    continue
                progress_col = active_cols.get("progress / run-rate") if active_cols else None
                status_col = active_cols.get("status") if active_cols else None
                if not (progress_col and status_col):
                    continue
                progress_txt = str(ws.cell(rr, progress_col).value or "").strip()
                if progress_txt.lower() in status_like_progress:
                    ws.cell(rr, progress_col).value = ""

        if ws.title == "Valuation":
            side_start, side_end = 15, min(max_col, 29)
            if ticker_txt == "GTX":
                for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        if isinstance(cell.value, str):
                            cell.value = (
                                cell.value
                                .replace("2025 year actuals", "FY2025 actuals")
                                .replace("2025 year adjusted", "FY2025 adjusted")
                                .replace("2025 year buybacks", "FY2025 buybacks")
                            )
            side_header_tokens = {
                "metric",
                "operating drivers",
                "thesis bridge",
                "bridge item",
                "output",
                "interpretation",
            }
            for rr in range(1, max_row + 1):
                dim = ws.row_dimensions[rr]
                if bool(dim.hidden):
                    continue
                for cc in range(side_start, side_end + 1):
                    cell = ws.cell(rr, cc)
                    raw_txt = str(cell.value or "")
                    if raw_txt.startswith("Guidance (As of") and "| Found:" in raw_txt:
                        header_txt, found_txt = raw_txt.split("| Found:", 1)
                        cell.value = header_txt.rstrip(" -| ")
                        found_txt = found_txt.strip()
                        if found_txt and found_txt.lower() != "none":
                            cell.comment = Comment(f"Found metrics: {found_txt}", "Codex")
                vals = [str(ws.cell(rr, cc).value or "").strip().lower() for cc in range(1, min(max_col, 29) + 1)]
                has_side_panel = side_start <= side_end and any(ws.cell(rr, cc).value not in (None, "") for cc in range(side_start, side_end + 1))
                col_a = vals[0] if vals else ""
                col_o = vals[14] if len(vals) >= 15 else ""
                col_o_blue = (
                    max_col >= 15
                    and str(ws.cell(rr, 15).fill.fgColor.rgb or "").upper().endswith(("5B9BD5", "6FA8DC"))
                    and bool(col_o)
                )
                side_header_match = col_o in side_header_tokens or col_o.startswith("guidance (as of")
                title_match = any(str(ws.cell(rr, cc).value or "").strip().lower() == "valuation" for cc in range(1, min(max_col, 14) + 1))
                has_main_content = any(ws.cell(rr, cc).value not in (None, "") for cc in range(1, min(max_col, 14) + 1))
                if title_match:
                    dim.height = 21.0
                    for cc in range(1, min(max_col, 14) + 1):
                        c = ws.cell(rr, cc)
                        if str(c.value or "").strip().lower() == "valuation":
                            c.font = Font(
                                name=c.font.name,
                                sz=18,
                                bold=True,
                                italic=c.font.italic,
                                color="FFFFFF",
                            )
                elif rr >= 6 and has_main_content:
                    dim.height = 19.5
                elif has_side_panel:
                    dim.height = 21.0 if (col_o_blue and side_header_match) else 19.5
                elif rr >= 6:
                    dim.height = 19.5
                else:
                    dim.height = 18.5
        elif ws.title == "BS_Segments":
            _clamp_rows(5, 18.0, 19.5)
        elif ws.title == "Operating_Drivers":
            if ticker_txt == "GTX":
                _clamp_rows(5, 19.5, 32.0)
            else:
                _clamp_rows(5, 19.5, 22.5)
        elif ws.title == "Quarter_Notes_UI":
            first_cell = str(ws.cell(1, 1).value or "").strip()
            second_cell = str(ws.cell(2, 1).value or "").strip()
            if "Quarter Notes" in first_cell and second_cell == "Quarter read":
                # Narrative Quarter_Notes_UI owns its row heights because body
                # rows contain wrapped multi-column reads.  The legacy compact
                # clamp is only safe for the old short-note layout.
                pass
            else:
                _standardize_quarter_notes_ui_categories(ws, ticker_txt)
                _clamp_rows(3, 19.5, 20.0)
        elif ws.title == "Promise_Progress_UI":
            if ticker_txt == "GTX":
                for row in ws.iter_rows(min_row=1, max_row=int(ws.max_row or 0), min_col=1, max_col=int(ws.max_column or 0)):
                    for cell in row:
                        if isinstance(cell.value, str) and "2025 year" in cell.value:
                            cell.value = cell.value.replace("2025 year", "FY2025")
                _clamp_rows(3, 26.0, 34.0)
                continue
            _standardize_promise_status_cells()
            _remove_empty_promise_revision_blocks(ws)
            max_row = int(ws.max_row or 0)
            max_col = int(ws.max_column or 0)
            _clamp_rows(3, 22.0, 26.0)
            _polish_promise_scorecard_layout(ws)
            _apply_source_backed_promise_mapping_overrides(wb, ticker_txt)
        elif ws.title == "Economics_Overlay" or ws.title.endswith("_Economics_Overlay"):
            for rr in range(1, max_row + 1):
                dim = ws.row_dimensions[rr]
                if bool(dim.hidden):
                    continue
                cur_f = float(dim.height or 21.0)
                row_fill = str(ws.cell(rr, 1).fill.fgColor.rgb or "")
                first_val = str(ws.cell(rr, 1).value or "").strip()
                row_text = " ".join(
                    str(ws.cell(rr, cc).value or "").strip()
                    for cc in range(1, min(max_col, 21) + 1)
                ).strip()
                if not row_text and cur_f <= 12.5:
                    dim.height = cur_f
                elif first_val.lower() == "electricity usage":
                    dim.height = 33.0
                elif first_val.lower() == "market inputs":
                    dim.height = 21.0
                elif row_fill.upper().endswith("5B9BD5") and first_val:
                    dim.height = 21.0
                    ws.cell(rr, 1).font = copy(ws.cell(rr, 1).font)
                    ws.cell(rr, 1).font = Font(
                        name=ws.cell(rr, 1).font.name,
                        sz=12,
                        bold=True,
                        italic=ws.cell(rr, 1).font.italic,
                        color="FFFFFF",
                    )
                elif cur_f > 24.0 and len(" ".join(str(ws.cell(rr, cc).value or "") for cc in range(1, min(max_col, 8) + 1)).strip()) < 80:
                    dim.height = 24.0
                elif cur_f > 42.0:
                    dim.height = 42.0
        elif ws.title.endswith("_Investment_Case") or ws.title == "ANF_Investment_Case":
            for rr in range(1, max_row + 1):
                dim = ws.row_dimensions[rr]
                if bool(dim.hidden):
                    continue
                cur_f = float(dim.height or 21.0)
                row_fill = str(ws.cell(rr, 1).fill.fgColor.rgb or "")
                first_val = str(ws.cell(rr, 1).value or "").strip()
                if row_fill.upper().endswith("5B9BD5") and first_val:
                    dim.height = 21.0
                    ws.cell(rr, 1).font = copy(ws.cell(rr, 1).font)
                    ws.cell(rr, 1).font = Font(
                        name=ws.cell(rr, 1).font.name,
                        sz=12,
                        bold=True,
                        italic=ws.cell(rr, 1).font.italic,
                        color="FFFFFF",
                    )
                elif cur_f < 21.0:
                    dim.height = 21.0
                elif cur_f > 42.0:
                    dim.height = 42.0
            _polish_investment_case_readability(ws)

    def _clean_guidance_normalized_sheet() -> None:
        if ticker_txt not in {"PBI", "GPRE"}:
            return
        for sheet_name in ("Guidance_Normalized", "Slides_Guidance"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            if int(ws.max_row or 0) < 2:
                continue
            headers = [str(ws.cell(1, cc).value or "").strip().lower() for cc in range(1, int(ws.max_column or 0) + 1)]
            metric_col = (headers.index("metric") + 1) if "metric" in headers else None
            horizon_col = (headers.index("horizon_label") + 1) if "horizon_label" in headers else None
            source_date_col = (headers.index("source_date") + 1) if "source_date" in headers else None
            line_col = (headers.index("line") + 1) if "line" in headers else None
            context_col = (headers.index("source_context") + 1) if "source_context" in headers else None
            source_col = (headers.index("source") + 1) if "source" in headers else None
            value_col = (headers.index("value") + 1) if "value" in headers else None
            low_col = (headers.index("low") + 1) if "low" in headers else None
            high_col = (headers.index("high") + 1) if "high" in headers else None
            unit_col = (headers.index("unit") + 1) if "unit" in headers else None
            basis_col = (headers.index("basis") + 1) if "basis" in headers else None
            rows_to_delete: List[int] = []
            for rr in range(2, int(ws.max_row or 0) + 1):
                vals = [ws.cell(rr, cc).value for cc in range(1, int(ws.max_column or 0) + 1)]
                blob = " | ".join(str(v) for v in vals if v is not None)
                metric_txt = str(ws.cell(rr, metric_col).value or "").strip() if metric_col else ""
                horizon_txt = str(ws.cell(rr, horizon_col).value or "").strip() if horizon_col else ""
                source_date_txt = str(ws.cell(rr, source_date_col).value or "").strip() if source_date_col else ""
                line_txt = str(ws.cell(rr, line_col).value or "").strip() if line_col else ""
                context_txt = str(ws.cell(rr, context_col).value or "").strip() if context_col else ""
                source_txt = str(ws.cell(rr, source_col).value or "").strip().lower() if source_col else ""
                value_txt = str(ws.cell(rr, value_col).value or "").strip() if value_col else ""
                low_txt = str(ws.cell(rr, low_col).value or "").strip() if low_col else ""
                high_txt = str(ws.cell(rr, high_col).value or "").strip() if high_col else ""
                unit_txt = str(ws.cell(rr, unit_col).value or "").strip().lower() if unit_col else ""
                basis_txt = str(ws.cell(rr, basis_col).value or "").strip().lower() if basis_col else ""
                blob_low = blob.lower()
                guidance_context = f"{line_txt} {context_txt}".lower()
                explicit_guidance_context = bool(
                    re.search(
                        r"\b(expects?|expected|outlook|guidance|forecast|anticipates?|currently expects|target range|annualized savings target|raised guidance|lowered guidance|maintained guidance)\b",
                        guidance_context,
                        re.I,
                    )
                )
                metric_low = metric_txt.lower()
                value_blob = " | ".join(x for x in (value_txt, low_txt, high_txt, line_txt, context_txt) if x)
                numeric_tokens = [
                    float(x.replace(",", ""))
                    for x in re.findall(r"(?<![A-Za-z])[-+]?\d+(?:,\d{3})*(?:\.\d+)?", value_blob)
                    if x not in {"2024", "2025", "2026", "2027"}
                ]

                def _bad_numeric_or_unit_shape() -> bool:
                    if "q2 commercial setup" in metric_low:
                        return True
                    if "competition alternative fuels" in blob_low:
                        return True
                    if "analyst" in blob_low and re.search(r"\b(question|asked|q&a)\b", blob_low, re.I):
                        return True
                    if re.search(r"\b(revenue|sales)\b", metric_low, re.I):
                        if re.search(r"\$\s*0\.\d+", value_blob):
                            return True
                        if "share" in unit_txt:
                            return True
                        if len(numeric_tokens) >= 2 and not re.search(r"(%|percent|bps|\$|\bm\b|\bbn\b|million|billion)", value_blob, re.I):
                            lo, hi = min(numeric_tokens), max(numeric_tokens)
                            if hi <= 20 and lo >= 0:
                                return True
                    if re.search(r"\beps\b", metric_low, re.I):
                        if unit_txt in {"%", "percent", "bps", "m shares", "shares"}:
                            return True
                        if len(numeric_tokens) >= 2:
                            lo, hi = min(numeric_tokens), max(numeric_tokens)
                            if lo < 2 and hi > 5:
                                return True
                    if re.search(r"\bcapex|capital expenditure", metric_low, re.I) and "shares" in unit_txt:
                        return True
                    if re.search(r"\bshares?\b", metric_low, re.I) and unit_txt in {"$", "$m", "%", "bps"}:
                        return True
                    if not unit_txt and not re.search(r"(%|percent|bps|\$|\bm\b|\bbn\b|million|billion|shares?)", value_blob, re.I):
                        if numeric_tokens and not explicit_guidance_context:
                            return True
                    return False

                actual_result_commentary = bool(
                    re.search(r"\b(reported|actual|was|were|increased|decreased|compared to|primarily affected)\b", blob_low)
                    and re.search(r"\b(revenue|net sales|capital expenditures|investing activities|adjusted ebitda|net income)\b", blob_low)
                    and not explicit_guidance_context
                )
                analyst_or_raw_fragment = bool(
                    re.search(r"\b(analyst|operator|questions?|q&a|asked)\b", blob_low)
                    or re.search(r"\b(can you comment|thank you for taking the questions|taking the questions)\b", blob_low)
                    or "?" in blob_low
                    or re.search(r"\b(other competition alternative fuels|competition alternative fuels)\b", blob_low)
                    or re.search(r"\b(html|webcast transcript|conference call transcript)\b", source_txt)
                    and metric_txt.lower() in {"other", "unknown", ""}
                )
                stale_or_unscoped_actual = bool(
                    re.search(r"\b(revenue|sales|capex|capital expenditures|investing activities|cash provided|cash used)\b", blob_low)
                    and re.search(r"\b(actual|reported|was|were|compared with|compared to|primarily affected|lower capital expenditures)\b", blob_low)
                    and not explicit_guidance_context
                )
                stock_comp_valuation_footnote = bool(
                    re.search(r"\bstock options?\b", blob_low)
                    and re.search(r"\bfair value\b|\bblack[- ]scholes\b|\bgrant date\b", blob_low)
                )
                debt_retirement_boilerplate = bool(
                    re.search(r"\b(convertible\s+notes?|notes?|debt|borrowings?)\b", blob_low)
                    and re.search(r"\b(maturit|matures?|retir(?:e|ing)|outstanding)\b", blob_low)
                    and not re.search(r"\b(guidance|outlook|target|range|forecast|guide)\b", guidance_context, re.I)
                )
                debt_boilerplate_not_guidance = bool(
                    re.search(r"\b(debt|revolver|abl|credit facility|maturit|borrowings?|notes?)\b", blob_low)
                    and (not explicit_guidance_context or debt_retirement_boilerplate)
                )
                raw_fragment_or_boilerplate = bool(
                    re.search(r"\binterim period results are not necessarily indicative\b", blob_low)
                    or re.search(r"\bto repeat sales customers and anticipate expanding it\b", blob_low)
                    or re.search(r"\bgood morning and thank you\b", blob_low)
                    or re.search(r"\basc(?:\s+topic)?\s*(?:606|842)\b", blob_low)
                    or re.search(r"\brevenue recognition\b", blob_low)
                    or re.search(r"\blease accounting(?:\s+system)?\b", blob_low)
                    or re.search(r"\bamended accounting guidance\b", blob_low)
                    or re.search(r"\baccounting standards? update\b", blob_low)
                    or re.search(r"\badopt(?:ed|ion of)\s+(?:new\s+)?accounting\b", blob_low)
                    or re.search(r"\byear;\s*updates\s+gaap\s+eps\b", blob_low)
                    or re.search(r"\bissued\s+\$?\s*500\s+million\b.*\b(retired|redeem|repay)\w*\b.*\bdebt\b", blob_low)
                    or re.search(r"\b(reclassifications|recurring in nature|unless otherwise noted)\b", blob_low)
                    or re.search(r"\b8217\b|\b8220\b|\b8221\b", blob_low)
                )
                qualitative_guidance_ok = bool(
                    metric_txt
                    and metric_txt.lower() not in {"other", "unknown", "nan", "none"}
                    and horizon_txt
                    and source_date_txt
                    and explicit_guidance_context
                    and (line_txt or context_txt)
                )
                noisy = (
                    "&#" in blob
                    or metric_txt.lower() in {"", "other", "nan", "none"}
                    or raw_fragment_or_boilerplate
                    or _bad_numeric_or_unit_shape()
                    or actual_result_commentary
                    or analyst_or_raw_fragment
                    or stale_or_unscoped_actual
                    or stock_comp_valuation_footnote
                    or debt_boilerplate_not_guidance
                    or (
                        "investing activities" in blob_low
                        and "capital expenditures" in blob_low
                        and not explicit_guidance_context
                    )
                    or (
                        sheet_name == "Guidance_Normalized"
                        and not horizon_txt
                    )
                    or (
                        not horizon_txt
                        and not explicit_guidance_context
                        and not re.search(r"\brun[- ]rate|savings target\b", guidance_context, re.I)
                    )
                    or (
                        not horizon_txt
                        and source_txt in {"financial_statement", "sec_filing", "annual_report"}
                        and not explicit_guidance_context
                    )
                    or not source_date_txt
                )
                if qualitative_guidance_ok and not (
                    raw_fragment_or_boilerplate
                    or analyst_or_raw_fragment
                    or actual_result_commentary
                    or stale_or_unscoped_actual
                    or debt_boilerplate_not_guidance
                    or stock_comp_valuation_footnote
                    or _bad_numeric_or_unit_shape()
                ):
                    noisy = False
                if noisy:
                    rows_to_delete.append(rr)
            for rr in reversed(rows_to_delete):
                ws.delete_rows(rr, 1)

    def _polish_needs_review_display() -> None:
        if "Needs_Review" not in wb.sheetnames:
            return
        ws = wb["Needs_Review"]
        if int(ws.max_row or 0) < 1:
            return
        headers = [str(ws.cell(1, cc).value or "").strip().lower() for cc in range(1, int(ws.max_column or 0) + 1)]
        width_by_header = {
            "priority": 18.0,
            "issue_family": 30.0,
            "severity": 12.0,
            "first_seen_q": 18.0,
            "last_seen_q": 18.0,
            "quarter_count": 14.0,
            "latest_message": 48.0,
            "recommended_action": 50.0,
            "source": 26.0,
            "quarter": 18.0,
        }
        date_headers = {"first_seen_q", "last_seen_q", "quarter", "date", "as_of"}
        for idx, header_name in enumerate(headers, start=1):
            if not header_name:
                continue
            letter = get_column_letter(idx)
            target_width = width_by_header.get(header_name)
            if target_width is None:
                normalized = header_name.replace("_", " ").strip().lower()
                if normalized.endswith(" seen q") or normalized.endswith(" date"):
                    target_width = 18.0
            if target_width is not None:
                ws.column_dimensions[letter].width = max(float(ws.column_dimensions[letter].width or 0), target_width)
            if header_name in date_headers or header_name.endswith("_q") or header_name.endswith("_date"):
                for date_rr in range(2, int(ws.max_row or 0) + 1):
                    ws.cell(date_rr, idx).number_format = "yyyy-mm-dd"
                    ws.cell(date_rr, idx).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if header_name in {"latest_message", "recommended_action"}:
                for msg_rr in range(2, int(ws.max_row or 0) + 1):
                    ws.cell(msg_rr, idx).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    def _downgrade_pbi_qsum_false_positive_rows() -> None:
        if ticker_txt not in {"PBI", "ANF"}:
            return
        for sheet_name in ("Needs_Review", "QA_Checks", "QA_Log"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            if int(ws.max_row or 0) < 2:
                continue
            headers = [str(ws.cell(1, cc).value or "").strip().lower() for cc in range(1, int(ws.max_column or 0) + 1)]
            col_map = {h: idx + 1 for idx, h in enumerate(headers) if h}
            if sheet_name == "Needs_Review":
                for header_name, cc in col_map.items():
                    normalized_header = header_name.replace("_", " ").strip().lower()
                    if (
                        normalized_header in {"first seen", "last seen", "first seen q", "last seen q", "date", "as of"}
                        or normalized_header.endswith(" seen q")
                        or normalized_header.endswith(" date")
                    ):
                        letter = get_column_letter(cc)
                        ws.column_dimensions[letter].width = max(float(ws.column_dimensions[letter].width or 0), 18.0)
                        for date_rr in range(2, int(ws.max_row or 0) + 1):
                            ws.cell(date_rr, cc).number_format = "yyyy-mm-dd"
            for rr in range(2, int(ws.max_row or 0) + 1):
                row_blob = " | ".join(
                    str(ws.cell(rr, cc).value or "")
                    for cc in range(1, int(ws.max_column or 0) + 1)
                ).lower()
                if "qsum_vs_fy" not in row_blob:
                    continue
                message_col = col_map.get("message") or col_map.get("latest_message")
                message = str(ws.cell(rr, message_col).value or "") if message_col else ""
                if "sum of 4 quarters vs fy fact" not in message.lower() and "sum of # quarters vs fy" not in message.lower():
                    continue
                for col_name in ("severity", "status"):
                    cc = col_map.get(col_name)
                    if cc and str(ws.cell(rr, cc).value or "").strip().lower() == "fail":
                        ws.cell(rr, cc).value = "warn"
                priority_col = col_map.get("priority")
                if priority_col and str(ws.cell(rr, priority_col).value or "").strip().upper() == "P1":
                    ws.cell(rr, priority_col).value = "P2"
                action_col = col_map.get("recommended_action")
                if action_col:
                    existing = str(ws.cell(rr, action_col).value or "").strip()
                    note = "Review as an older fiscal/YTD mapping limitation; current core valuation outputs are not blocked."
                    ws.cell(rr, action_col).value = (
                        existing if note in existing else (f"{existing} {note}".strip() if existing else note)
                    )
                review_status_col = col_map.get("review_status")
                if review_status_col:
                    ws.cell(rr, review_status_col).value = "Watch"

    _clean_guidance_normalized_sheet()
    _polish_needs_review_display()
    _downgrade_pbi_qsum_false_positive_rows()
