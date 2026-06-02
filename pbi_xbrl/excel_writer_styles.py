"""Pure style bundle factories shared by workbook writer surfaces."""
from __future__ import annotations

from typing import Any, Dict

from openpyxl.styles import Border, Font, PatternFill, Side


def get_analysis_sheet_style_bundle(header_size: float, font_size: float) -> Dict[str, Any]:
    border_color = "AAB7C4"
    text_dark = "1F1F1F"
    text_muted = "5E6F82"
    accent_text = "274B78"
    thin = Side(style="thin", color=border_color)
    return {
        "title_fill": PatternFill("solid", fgColor="6FA8DC"),
        "section_fill": PatternFill("solid", fgColor="EDF4FA"),
        "header_fill": PatternFill("solid", fgColor="EAF3FB"),
        "neutral_fill": PatternFill("solid", fgColor="F7F9FC"),
        "neutral_fill_alt": PatternFill("solid", fgColor="FFFFFF"),
        "input_fill": PatternFill("solid", fgColor="FFF2CC"),
        "text_dark": text_dark,
        "text_muted": text_muted,
        "accent_text": accent_text,
        "border_color": border_color,
        "thin_side": thin,
        "thin_border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "title_font": Font(bold=True, size=15, color="FFFFFF"),
        "bold_font": Font(bold=True, size=header_size, color=text_dark),
        "norm_font": Font(size=font_size, color=text_dark),
        "muted_font": Font(size=font_size, color=text_muted),
    }


def valuation_side_panel_style_bundle() -> Dict[str, Any]:
    """Shared ANF-style Valuation side-panel styling for all tickers."""
    soft_side = Side(style="thin", color="D9E2EA")
    return {
        "section_fill": PatternFill("solid", fgColor="6FA8DC"),
        "header_fill": PatternFill("solid", fgColor="EAF3FB"),
        "neutral_fill": PatternFill("solid", fgColor="FFFFFF"),
        "neutral_alt_fill": PatternFill("solid", fgColor="F7FAFC"),
        "input_fill": PatternFill("solid", fgColor="FFF2CC"),
        "thin_border": Border(left=soft_side, right=soft_side, top=soft_side, bottom=soft_side),
        "title_font": Font(bold=True, size=12, color="FFFFFF"),
        "header_font": Font(bold=True, size=12, color="1F2933"),
        "body_font": Font(size=12, color="1F2933"),
        "input_font": Font(size=12, color="0000FF"),
    }
