"""Pure Summary projections for filing freshness and post-quarter effects."""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SOURCE_FILING_FRESHNESS_COLUMNS = (
    "ticker",
    "latest_reported_quarter",
    "latest_reported_filing_type",
    "latest_reported_filing_accession",
    "latest_reported_filing_date",
    "latest_reported_downloaded_at",
    "latest_additional_filing_type",
    "latest_additional_filing_accession",
    "latest_additional_filing_date",
    "latest_additional_downloaded_at",
    "event_type",
    "used_in_workbook",
    "used_surfaces",
    "source_path_exists",
)

POST_QUARTER_CURRENT_EFFECT_COLUMNS = (
    "ticker",
    "event_date",
    "filing_date",
    "area",
    "reported_quarter_anchor",
    "reported_value",
    "current_overlay_value",
    "change",
    "unit",
    "confidence_treatment",
    "historical_treatment",
    "valuation_treatment",
    "source_document_accession",
)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _quarter_label(value: Any) -> str:
    timestamp = pd.to_datetime(value, errors="coerce")
    if pd.isna(timestamp):
        return ""
    return f"{timestamp.year}-Q{timestamp.quarter}"


def _date_text(value: Any) -> str:
    timestamp = pd.to_datetime(value, errors="coerce")
    if pd.isna(timestamp):
        return ""
    return timestamp.strftime("%Y-%m-%d")


def _path_from_row(row: pd.Series) -> str:
    for column in (
        "doc",
        "source_path",
        "source_local_path",
        "materialized_path",
        "local_path",
    ):
        value = _clean_text(row.get(column))
        if value:
            return value
    return ""


def _downloaded_at(row: pd.Series, source_path: str) -> str:
    for column in ("downloaded_at", "download_timestamp"):
        value = _clean_text(row.get(column))
        if value:
            return value
    if source_path:
        path = Path(source_path)
        try:
            return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()
        except OSError:
            return ""
    return ""


def _latest_reported_metadata(
    *,
    latest_quarter: pd.Timestamp | None,
    audit: Any,
    manifest_df: Any,
) -> dict[str, str]:
    candidates: list[pd.DataFrame] = []
    if isinstance(audit, pd.DataFrame) and not audit.empty:
        frame = audit.copy()
        frame["_metadata_source_priority"] = 2
        quarter_column = "quarter" if "quarter" in frame.columns else "report_date"
        if quarter_column in frame.columns:
            frame["_quarter"] = pd.to_datetime(frame[quarter_column], errors="coerce")
            if latest_quarter is not None:
                frame = frame[
                    frame["_quarter"].dt.to_period("Q")
                    == latest_quarter.to_period("Q")
                ]
        if not frame.empty:
            candidates.append(frame)
    if isinstance(manifest_df, pd.DataFrame) and not manifest_df.empty:
        frame = manifest_df.copy()
        frame["_metadata_source_priority"] = 1
        quarter_column = "reportDate" if "reportDate" in frame.columns else "report_date"
        if quarter_column in frame.columns:
            frame["_quarter"] = pd.to_datetime(frame[quarter_column], errors="coerce")
            if latest_quarter is not None:
                frame = frame[
                    frame["_quarter"].dt.to_period("Q")
                    == latest_quarter.to_period("Q")
                ]
        if not frame.empty:
            candidates.append(frame)
    if not candidates:
        return {
            "filing_type": "",
            "accession": "",
            "filing_date": "",
            "downloaded_at": "",
            "source_path": "",
        }

    frame = pd.concat(candidates, ignore_index=True, sort=False)
    filing_column = next(
        (column for column in ("filed", "filedDate", "filing_date") if column in frame.columns),
        "",
    )
    frame["_filing_date"] = (
        pd.to_datetime(frame[filing_column], errors="coerce")
        if filing_column
        else pd.NaT
    )
    frame["_metadata_score"] = frame.apply(
        lambda row: sum(
            bool(_clean_text(row.get(column)))
            for column in (
                "form",
                "accn",
                "accession",
                "accessionNumber",
                "doc",
                "source_local_path",
                "materialized_path",
                "downloaded_at",
            )
        ),
        axis=1,
    )
    frame = frame.sort_values(
        ["_filing_date", "_metadata_score", "_metadata_source_priority"],
        na_position="first",
    )
    row = frame.iloc[-1]
    source_path = _path_from_row(row)
    accession = ""
    for column in ("accn", "accession", "accessionNumber"):
        value = _clean_text(row.get(column))
        if value:
            accession = value
            break
    return {
        "filing_type": _clean_text(row.get("form")),
        "accession": accession,
        "filing_date": _date_text(row.get(filing_column)) if filing_column else "",
        "downloaded_at": _downloaded_at(row, source_path),
        "source_path": source_path,
    }


def build_source_filing_freshness(
    *,
    ticker: str,
    hist: Any,
    audit: Any,
    manifest_df: Any,
    post_quarter_events: Any,
    source_roots: Any = (),
) -> pd.DataFrame:
    ticker_key = str(ticker or "").strip().upper()
    latest_quarter: pd.Timestamp | None = None
    if isinstance(hist, pd.DataFrame) and not hist.empty and "quarter" in hist.columns:
        quarters = pd.to_datetime(hist["quarter"], errors="coerce").dropna()
        if not quarters.empty:
            latest_quarter = pd.Timestamp(quarters.max())
    reported = _latest_reported_metadata(
        latest_quarter=latest_quarter,
        audit=audit,
        manifest_df=manifest_df,
    )
    if not reported["downloaded_at"] and reported["accession"]:
        accession_token = "".join(
            character
            for character in reported["accession"]
            if character.isdigit()
        )
        source_candidates: list[Path] = []
        for root_value in source_roots or ():
            root = Path(root_value).expanduser()
            if not root.exists():
                continue
            source_candidates.extend(
                path
                for path in root.rglob(f"*{accession_token}*")
                if path.is_file()
            )
        if source_candidates:
            latest_source = max(
                source_candidates,
                key=lambda path: path.stat().st_mtime,
            )
            reported["source_path"] = str(latest_source)
            reported["downloaded_at"] = (
                datetime.fromtimestamp(
                    latest_source.stat().st_mtime,
                    tz=timezone.utc,
                ).isoformat()
                + " (filesystem fallback)"
            )

    events = (
        post_quarter_events.copy()
        if isinstance(post_quarter_events, pd.DataFrame)
        else pd.DataFrame()
    )
    if not events.empty and "ticker" in events.columns:
        events = events[events["ticker"].astype(str).str.upper().eq(ticker_key)]
    event = None
    if not events.empty:
        order = pd.to_datetime(events.get("filing_date"), errors="coerce")
        events = events.assign(_filing_order=order).sort_values(
            "_filing_order",
            na_position="last",
        )
        event = events.iloc[-1]

    event_type_map = {
        "refinancing_redemption": "refinancing/redemption",
        "warrant_dilution": "warrant dilution",
    }
    if event is None:
        additional_type = "None newer / no model-relevant post-quarter event"
        additional_accession = ""
        additional_filing_date = ""
        additional_downloaded_at = ""
        event_type = ""
        used_in_workbook = "No"
        used_surfaces = ""
        source_path_exists = "Yes" if reported["source_path"] and Path(reported["source_path"]).exists() else "No"
    else:
        additional_type = str(event.get("filing_type") or "").strip()
        additional_accession = str(event.get("accession") or "").strip()
        additional_filing_date = _date_text(event.get("filing_date"))
        additional_downloaded_at = str(event.get("downloaded_at") or "").strip()
        event_type_raw = str(event.get("event_type") or "").strip()
        event_type = event_type_map.get(event_type_raw, event_type_raw.replace("_", " "))
        used_in_workbook = str(event.get("used_in_workbook") or "Review").strip()
        used_surfaces = str(event.get("used_surfaces") or "").strip()
        source_path_exists = "Yes" if bool(event.get("source_path_exists")) else "No"

    row = {
        "ticker": ticker_key,
        "latest_reported_quarter": _quarter_label(latest_quarter),
        "latest_reported_filing_type": reported["filing_type"],
        "latest_reported_filing_accession": reported["accession"],
        "latest_reported_filing_date": reported["filing_date"],
        "latest_reported_downloaded_at": reported["downloaded_at"],
        "latest_additional_filing_type": additional_type,
        "latest_additional_filing_accession": additional_accession,
        "latest_additional_filing_date": additional_filing_date,
        "latest_additional_downloaded_at": additional_downloaded_at,
        "event_type": event_type,
        "used_in_workbook": used_in_workbook,
        "used_surfaces": used_surfaces,
        "source_path_exists": source_path_exists,
    }
    return pd.DataFrame([row], columns=SOURCE_FILING_FRESHNESS_COLUMNS)


def _source_label(event: pd.Series) -> str:
    documents = str(event.get("source_documents") or "").strip()
    accession = str(event.get("accession") or "").strip()
    if documents and accession:
        return f"{documents} | accession {accession}"
    return documents or accession


def build_post_quarter_current_effects(
    post_quarter_events: Any,
) -> pd.DataFrame:
    if not isinstance(post_quarter_events, pd.DataFrame) or post_quarter_events.empty:
        return pd.DataFrame(columns=POST_QUARTER_CURRENT_EFFECT_COLUMNS)
    rows: list[dict[str, Any]] = []
    for _, event in post_quarter_events.iterrows():
        common = {
            "ticker": str(event.get("ticker") or "").upper(),
            "event_date": _date_text(event.get("event_date")),
            "filing_date": _date_text(event.get("filing_date")),
            "reported_quarter_anchor": str(event.get("reported_quarter_anchor") or ""),
            "source_document_accession": _source_label(event),
        }
        event_type = str(event.get("event_type") or "")
        if event_type == "refinancing_redemption":
            principal_redeemed_m = float(event["principal_redeemed"]) / 1e6
            incremental_m = float(event["incremental_term_loan"]) / 1e6
            term_total_m = float(event["term_loan_total"]) / 1e6
            gross_delta_m = float(event["gross_principal_delta"]) / 1e6
            historical = "History_Q unchanged; Debt_Profile unchanged; Debt_Tranches_Latest unchanged"
            rows.extend(
                [
                    {
                        **common,
                        "area": "2027 Senior Notes",
                        "reported_value": principal_redeemed_m,
                        "current_overlay_value": 0.0,
                        "change": -principal_redeemed_m,
                        "unit": "$m principal",
                        "confidence_treatment": "Source-backed",
                        "historical_treatment": historical,
                        "valuation_treatment": "Current Debt Detail updated",
                    },
                    {
                        **common,
                        "area": "Term Loan A",
                        "reported_value": term_total_m - incremental_m,
                        "current_overlay_value": term_total_m,
                        "change": incremental_m,
                        "unit": "$m principal",
                        "confidence_treatment": "Source-backed",
                        "historical_treatment": historical,
                        "valuation_treatment": "Current Debt Detail updated",
                    },
                    {
                        **common,
                        "area": "Gross principal debt",
                        "reported_value": "",
                        "current_overlay_value": "",
                        "change": gross_delta_m,
                        "unit": "$m principal",
                        "confidence_treatment": "Source-backed",
                        "historical_treatment": historical,
                        "valuation_treatment": "Gross principal only",
                    },
                    {
                        **common,
                        "area": "Cash / net debt",
                        "reported_value": "Reported Q1 unchanged",
                        "current_overlay_value": "Unresolved / manual review",
                        "change": "",
                        "unit": "$m",
                        "confidence_treatment": "Partial / unresolved",
                        "historical_treatment": historical,
                        "valuation_treatment": "No auto net-debt adjustment",
                    },
                    {
                        **common,
                        "area": "Next scheduled maturity",
                        "reported_value": "March 2027",
                        "current_overlay_value": str(event.get("next_scheduled_maturity") or ""),
                        "change": "2027 maturity removed",
                        "unit": "date",
                        "confidence_treatment": "Source-backed",
                        "historical_treatment": historical,
                        "valuation_treatment": "Current Debt Detail updated",
                    },
                    {
                        **common,
                        "area": "Term Loan A maturity",
                        "reported_value": "May 18, 2031",
                        "current_overlay_value": "May 18, 2031",
                        "change": "Unchanged",
                        "unit": "date",
                        "confidence_treatment": "Source-backed",
                        "historical_treatment": historical,
                        "valuation_treatment": "Current Debt Detail updated",
                    },
                ]
            )
        elif event_type == "warrant_dilution":
            source_backed = "Source-backed"
            historical = "Shares/EPS unchanged"
            rows.extend(
                [
                    {
                        **common,
                        "area": "Warrants issued",
                        "reported_value": 0.0,
                        "current_overlay_value": float(event["warrants_issued"]),
                        "change": float(event["warrants_issued"]),
                        "unit": "warrants",
                        "confidence_treatment": source_backed,
                        "historical_treatment": historical,
                        "valuation_treatment": "Disclosure only",
                    },
                    {
                        **common,
                        "area": "Potential common shares issuable max",
                        "reported_value": 0.0,
                        "current_overlay_value": float(event["potential_common_shares_issuable_max"]),
                        "change": float(event["potential_common_shares_issuable_max"]),
                        "unit": "shares",
                        "confidence_treatment": source_backed,
                        "historical_treatment": historical,
                        "valuation_treatment": "Full-dilution sensitivity",
                    },
                    {
                        **common,
                        "area": "Valuation full-dilution overlay",
                        "reported_value": 0.0,
                        "current_overlay_value": 0.550,
                        "change": 0.550,
                        "unit": "million shares",
                        "confidence_treatment": source_backed,
                        "historical_treatment": historical,
                        "valuation_treatment": "Full-dilution sensitivity",
                    },
                    {
                        **common,
                        "area": "Exercise price",
                        "reported_value": "",
                        "current_overlay_value": float(event["exercise_price"]),
                        "change": "",
                        "unit": "$/share",
                        "confidence_treatment": source_backed,
                        "historical_treatment": historical,
                        "valuation_treatment": "Disclosure only",
                    },
                    {
                        **common,
                        "area": "Expiration",
                        "reported_value": "",
                        "current_overlay_value": str(event.get("expiration_date") or ""),
                        "change": "",
                        "unit": "date",
                        "confidence_treatment": source_backed,
                        "historical_treatment": historical,
                        "valuation_treatment": "Disclosure only",
                    },
                    {
                        **common,
                        "area": "Reported shares / EPS",
                        "reported_value": "Reported Q1",
                        "current_overlay_value": "Unchanged",
                        "change": 0.0,
                        "unit": "shares / EPS",
                        "confidence_treatment": source_backed,
                        "historical_treatment": historical,
                        "valuation_treatment": "Full-dilution sensitivity only",
                    },
                ]
            )
    return pd.DataFrame(rows, columns=POST_QUARTER_CURRENT_EFFECT_COLUMNS)


def append_summary_freshness_sections(
    *,
    ws: Any,
    start_row: int,
    source_filing_freshness: pd.DataFrame,
    post_quarter_current_effects: pd.DataFrame,
    font_size: int,
    header_size: int,
) -> int:
    """Append the two user-facing Summary tables without deriving new facts."""

    section_fill = PatternFill("solid", fgColor="5B9BD5")
    header_fill = PatternFill("solid", fgColor="4472C4")
    row_fills = (
        PatternFill("solid", fgColor="D9EAF7"),
        PatternFill("solid", fgColor="EDF4FB"),
    )
    thin = Side(style="thin", color="AAB7C4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _section(title: str, row: int, end_column: int) -> int:
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=end_column,
        )
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=header_size, color="FFFFFF")
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in range(1, end_column + 1):
            ws.cell(row=row, column=column).fill = section_fill
            ws.cell(row=row, column=column).border = border
        ws.row_dimensions[row].height = 24
        return row + 1

    def _table(
        frame: pd.DataFrame,
        *,
        row: int,
        columns: tuple[str, ...],
        labels: tuple[str, ...],
        group_labels: tuple[tuple[str, int, int], ...],
        no_data_text: str,
    ) -> int:
        for label, start_column, end_column in group_labels:
            if end_column > start_column:
                ws.merge_cells(
                    start_row=row,
                    start_column=start_column,
                    end_row=row,
                    end_column=end_column,
                )
            cell = ws.cell(row=row, column=start_column, value=label)
            cell.font = Font(bold=True, size=font_size, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border
            for column_index in range(start_column, end_column + 1):
                group_cell = ws.cell(row=row, column=column_index)
                group_cell.fill = header_fill
                group_cell.border = border
        ws.row_dimensions[row].height = 22
        row += 1
        for column_index, label in enumerate(labels, start=1):
            cell = ws.cell(row=row, column=column_index, value=label)
            cell.font = Font(bold=True, size=font_size, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border
        ws.row_dimensions[row].height = 30
        row += 1
        if frame.empty:
            ws.merge_cells(
                start_row=row,
                start_column=1,
                end_row=row,
                end_column=len(columns),
            )
            ws.cell(row=row, column=1, value=no_data_text)
            ws.cell(row=row, column=1).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )
            for column_index in range(1, len(columns) + 1):
                ws.cell(row=row, column=column_index).fill = row_fills[0]
                ws.cell(row=row, column=column_index).border = border
            return row + 1
        for record_index, record in enumerate(frame.to_dict("records")):
            fill = row_fills[record_index % 2]
            for column_index, column in enumerate(columns, start=1):
                value = record.get(column)
                if value is None or (not isinstance(value, str) and pd.isna(value)):
                    value = ""
                cell = ws.cell(row=row, column=column_index, value=value)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(
                    horizontal=(
                        "right"
                        if column
                        in {
                            "reported_value",
                            "current_overlay_value",
                            "change",
                        }
                        and isinstance(value, (int, float))
                        else "left"
                    ),
                    vertical="top",
                    wrap_text=True,
                )
                if column in {
                    "latest_reported_filing_date",
                    "latest_additional_filing_date",
                    "event_date",
                    "filing_date",
                }:
                    cell.number_format = "yyyy-mm-dd"
                if (
                    column
                    in {
                        "reported_value",
                        "current_overlay_value",
                        "change",
                    }
                    and isinstance(value, (int, float))
                ):
                    cell.number_format = "#,##0.000"
            ws.row_dimensions[row].height = 44
            row += 1
        return row

    freshness_columns = (
        "latest_reported_quarter",
        "latest_reported_filing_type",
        "latest_reported_filing_accession",
        "latest_reported_filing_date",
        "latest_reported_downloaded_at",
        "latest_additional_filing_type",
        "latest_additional_filing_accession",
        "latest_additional_filing_date",
        "latest_additional_downloaded_at",
        "event_type",
        "used_in_workbook",
        "used_surfaces",
        "source_path_exists",
        "ticker",
    )
    freshness_labels = (
        "Quarter",
        "Type",
        "Accession",
        "Filed",
        "Downloaded",
        "Type",
        "Accession",
        "Filed",
        "Downloaded",
        "Category",
        "Used?",
        "Surfaces",
        "Source ok?",
        "Ticker",
    )
    freshness_groups = (
        ("Reported filing", 1, 5),
        ("Post-quarter / current filing", 6, 9),
        ("Workbook use", 10, 13),
        ("Reference", 14, 14),
    )
    effects_columns = (
        "event_date",
        "filing_date",
        "area",
        "reported_quarter_anchor",
        "reported_value",
        "current_overlay_value",
        "change",
        "unit",
        "confidence_treatment",
        "historical_treatment",
        "valuation_treatment",
        "source_document_accession",
        "ticker",
    )
    effects_labels = (
        "Event date",
        "Filing date",
        "Category",
        "Reported quarter",
        "Reported value",
        "Current / overlay value",
        "Change",
        "Unit",
        "Confidence",
        "Historical handling",
        "Valuation handling",
        "Source / accession",
        "Ticker",
    )
    effects_groups = (
        ("Dates", 1, 2),
        ("Event", 3, 4),
        ("Values", 5, 8),
        ("Handling / control", 9, 11),
        ("Evidence", 12, 12),
        ("Reference", 13, 13),
    )

    row = start_row + 1
    row = _section("Source / Filing Freshness", row, len(freshness_columns))
    row = _table(
        source_filing_freshness,
        row=row,
        columns=freshness_columns,
        labels=freshness_labels,
        group_labels=freshness_groups,
        no_data_text="No filing freshness record available.",
    )
    row += 1
    row = _section("Post-quarter / Current Effects", row, len(effects_columns))
    row = _table(
        post_quarter_current_effects,
        row=row,
        columns=effects_columns,
        labels=effects_labels,
        group_labels=effects_groups,
        no_data_text="None newer / no model-relevant post-quarter event",
    )

    widths = {
        "A": 18,
        "B": 16,
        "C": 28,
        "D": 18,
        "E": 24,
        "F": 24,
        "G": 26,
        "H": 16,
        "I": 24,
        "J": 26,
        "K": 26,
        "L": 38,
        "M": 14,
        "N": 10,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = max(
            float(ws.column_dimensions[column].width or 0),
            width,
        )
    return row
