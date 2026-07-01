from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from typing import Any, Callable, Dict, List, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .excel_writer_summary_freshness import append_summary_freshness_sections


@dataclass(frozen=True)
class SummarySheetRenderDeps:
    wb: Any
    font_size: int
    header_size: int
    set_cell_comment: Callable[..., None]
    normalize_text: Callable[..., str]
    estimate_wrapped_line_count: Callable[..., int]
    estimate_wrapped_row_height: Callable[..., float]
    source_filing_freshness: Any
    post_quarter_current_effects: Any


def write_summary_sheet(
    deps: SummarySheetRenderDeps,
    df: Any,
) -> None:
    wb = deps.wb
    font_size = deps.font_size
    header_size = deps.header_size
    _set_cell_comment_local = deps.set_cell_comment
    glx_normalize_text = deps.normalize_text
    _estimate_wrapped_line_count = deps.estimate_wrapped_line_count
    _estimate_wrapped_row_height = deps.estimate_wrapped_row_height
    source_filing_freshness = deps.source_filing_freshness
    post_quarter_current_effects = deps.post_quarter_current_effects

    ws = wb.create_sheet(title="SUMMARY")
    if df is None or df.empty:
        ws["A1"] = "No data."
        return

    ws.sheet_format.defaultRowHeight = 18
    ws.sheet_view.zoomScale = 110
    ws.freeze_panes = "A2"
    section_fill = PatternFill("solid", fgColor="5B9BD5")
    row_fill_a = PatternFill("solid", fgColor="D9EAF7")
    row_fill_b = PatternFill("solid", fgColor="EDF4FB")
    thick_black = Side(style="thick", color="5E6F82")
    thin_gray = Side(style="thin", color="AAB7C4")
    overview_last_col = 6
    narrative_last_col = 6

    def _px_to_width(px: float) -> float:
        try:
            p = float(px)
        except Exception:
            p = 100.0
        return max(1.0, round((p - 5.0) / 7.0, 2))

    def _summary_unit_from_note(note: Any) -> str:
        txt = str(note or "").strip()
        if not txt:
            return ""
        low = txt.lower()
        if low == "date":
            return ""
        if low.startswith("units:"):
            return txt.split(":", 1)[1].strip().rstrip(".")
        if txt in {"$m", "%", "$/share", "x", "bps"}:
            return txt
        return ""

    def _summary_note_comment(note: Any) -> str:
        txt = str(note or "").strip()
        if not txt:
            return ""
        if txt.lower() == "date":
            return ""
        if txt.startswith("Units:"):
            return ""
        if txt in {"$m", "%", "$/share", "x", "bps"}:
            return ""
        return txt

    def _add_comment(cell, note_text: str) -> None:
        txt = str(note_text or "").strip()
        if not txt:
            return
        try:
            _set_cell_comment_local(cell, txt)
        except Exception:
            pass

    def _write_section_header(
        row_idx: int,
        title: str,
        title_size: float = 15.0,
        fill_color: str = "5B9BD5",
        font_color: str = "FFFFFF",
    ) -> int:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=overview_last_col)
        cell = ws.cell(row=row_idx, column=1, value=title)
        cell.font = Font(bold=True, size=title_size, color=font_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=fill_color)
        ws.row_dimensions[row_idx].height = 24
        for cc in range(1, overview_last_col + 1):
            ws.cell(row=row_idx, column=cc).fill = PatternFill("solid", fgColor=fill_color)
            ws.cell(row=row_idx, column=cc).border = Border(
                top=thick_black,
                left=thin_gray,
                right=thin_gray,
                bottom=thin_gray,
            )
        return row_idx + 1

    def _normalized_value_text(value: Any) -> str:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value)

    def _split_overview_value_lines(value_txt: str) -> List[str]:
        def _clean_overview_line(line_txt: str) -> str:
            s = glx_normalize_text(str(line_txt or ""))
            if s.startswith("- "):
                s = f"• {s[2:].strip()}"
            if re.search(r"[.!?%]$|\d$", s):
                return s
            last_word = re.sub(r"[^A-Za-z0-9%]+", "", s.split()[-1]) if s.split() else ""
            last_word_lc = last_word.lower()
            if len(s) > 100:
                cut_candidates = [
                    s.rfind(". "),
                    s.rfind("; "),
                    s.rfind(": "),
                    s.rfind(", "),
                ]
                cut_idx = max(cut_candidates)
                if cut_idx >= int(len(s) * 0.55):
                    trimmed = s[: cut_idx + 1].rstrip(" ,;:-")
                    if trimmed:
                        return trimmed
            if last_word_lc in {"around", "and", "the", "of", "to", "for", "in", "on", "with", "from", "by"}:
                cut_candidates = [s.rfind("; "), s.rfind(", "), s.rfind(". ")]
                cut_idx = max(cut_candidates)
                if cut_idx >= int(len(s) * 0.45):
                    trimmed = s[: cut_idx + 1].rstrip(" ,;:-")
                    if trimmed:
                        return trimmed
            parts = s.split()
            while parts:
                tail = re.sub(r"[^A-Za-z0-9%]+", "", parts[-1]).lower()
                if len(tail) <= 3:
                    parts.pop()
                    continue
                if len(tail) <= 4 and tail.isalpha() and tail == last_word_lc:
                    parts.pop()
                    continue
                break
            while parts and re.sub(r"[^A-Za-z0-9%]+", "", parts[-1]).lower() in {"and", "the", "of", "to", "for", "in", "on", "with", "from", "by"}:
                parts.pop()
            s = " ".join(parts).rstrip(" ,;:-")
            if s and not re.search(r"[.!?%]$|\d$", s):
                sentence_cut = max(s.rfind(". "), s.rfind("! "), s.rfind("? "))
                if sentence_cut >= int(len(s) * 0.55):
                    trimmed = s[: sentence_cut + 1].rstrip(" ,;:-")
                    if trimmed:
                        return trimmed
                clause_cut = max(s.rfind("; "), s.rfind(": "), s.rfind(", "))
                if clause_cut >= int(len(s) * 0.65):
                    trimmed = s[: clause_cut + 1].rstrip(" ,;:-")
                    if trimmed:
                        return trimmed
            return s or glx_normalize_text(str(line_txt or ""))

        parts: List[str] = []
        raw_parts = [str(p or "").strip() for p in str(value_txt or "").split("\n")]
        for part in raw_parts:
            if not part:
                continue
            parts.append(_clean_overview_line(part))
        return parts or ["N/A"]

    def _prefix_is_label_like(prefix_txt: str, delimiter: str) -> bool:
        prefix = str(prefix_txt or "").strip()
        if not prefix:
            return False
        core = prefix.rstrip(":;").strip()
        if not core:
            return False
        words = [w for w in re.split(r"\s+", core) if w]
        if delimiter == ":":
            return len(words) <= 10 and len(core) <= 90
        labelish_terms = {
            "advantage",
            "dependencies",
            "dependency",
            "segment",
            "segments",
            "model",
            "revenue",
            "risk",
            "spread",
            "sensitivity",
            "policy",
            "technology",
            "execution",
        }
        return len(words) <= 7 and (
            "/" in core
            or any(term in core.lower() for term in labelish_terms)
        )

    def _split_overview_prefix(line_txt: str) -> Tuple[str, str]:
        s = str(line_txt or "").strip()
        if not s:
            return "", ""
        bullet_match = re.match(r"^(•\s*)(.*)$", s)
        bullet_txt = ""
        core_txt = s
        if bullet_match:
            bullet_txt, core_txt = bullet_match.groups()
            core_txt = core_txt.strip()
        rich_match = re.match(r"^(.{1,90}?)([:;])(\s*)(.+)$", core_txt)
        if not rich_match:
            return "", s
        lead_core, delimiter, spacer_txt, rest_txt = rich_match.groups()
        lead_txt = f"{lead_core}{delimiter}"
        if not _prefix_is_label_like(lead_txt, delimiter):
            return "", s
        if not rest_txt.strip():
            return "", s
        lead_display = f"{bullet_txt}{lead_txt}".strip()
        rest_display = f"{spacer_txt}{rest_txt}".strip()
        return lead_display, rest_display

    def _fill_row(row_idx: int, fill: PatternFill, end_col: int = overview_last_col) -> None:
        for cc in range(1, end_col + 1):
            ws.cell(row=row_idx, column=cc).fill = fill

    def _top_black_line(row_idx: int, end_col: int = overview_last_col) -> None:
        for cc in range(1, end_col + 1):
            cell = ws.cell(row=row_idx, column=cc)
            cell.border = Border(
                top=thick_black,
                left=thin_gray,
                right=thin_gray,
                bottom=thin_gray,
            )

    def _summary_narrative_row_height(text: str, col_chars: float) -> float:
        est_lines = _estimate_wrapped_line_count(
            text,
            col_chars,
            min_lines=1,
            max_lines=12,
        )
        row_h = _estimate_wrapped_row_height(
            text,
            col_chars,
            20,
            14,
            min_lines=1,
            max_lines=12,
        )
        if est_lines > 2.15:
            row_h += 4.0
        if est_lines > 3.15:
            row_h += 4.0
        if est_lines > 4.15:
            row_h += 2.0
        return min(96.0, max(24.0, row_h))

    rows = df.to_dict("records")
    company_rows = [r for r in rows if str(r.get("Section") or "").strip().lower() == "company overview"]
    metric_rows = [
        r
        for r in rows
        if str(r.get("Section") or "").strip()
        and str(r.get("Section") or "").strip().lower() not in {"company overview", "qa"}
    ]

    overview_items: List[Dict[str, str]] = []
    current_item: Dict[str, str] | None = None
    for rec in company_rows:
        metric_txt = str(rec.get("Metric") or "").strip()
        value_txt = _normalized_value_text(rec.get("Value"))
        note_txt = str(rec.get("Note") or "").strip()
        if metric_txt:
            if current_item is not None:
                overview_items.append(current_item)
            current_item = {"metric": metric_txt, "value": "", "note": note_txt}
        if current_item is None:
            continue
        if value_txt:
            if current_item["value"]:
                current_item["value"] = f"{current_item['value']}\n{value_txt}"
            else:
                current_item["value"] = value_txt
        if not current_item.get("note") and note_txt:
            current_item["note"] = note_txt
    if current_item is not None:
        overview_items.append(current_item)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    merged_text_width = sum(
        float(ws.column_dimensions[get_column_letter(col)].width or 0.0) for col in range(1, narrative_last_col + 1)
    ) or _px_to_width(714)

    row_idx = _write_section_header(1, "Company Overview", 15.0)
    overview_band_idx = 0
    for item_idx, item in enumerate(overview_items):
        metric_txt = str(item.get("metric") or "").strip()
        value_txt = str(item.get("value") or "").strip() or "N/A"
        note_txt = str(item.get("note") or "").strip()
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=narrative_last_col)
        label_cell = ws.cell(row=row_idx, column=1, value=metric_txt)
        label_cell.font = Font(bold=True, size=header_size)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        _fill_row(row_idx, copy(section_fill), narrative_last_col)
        _top_black_line(row_idx, narrative_last_col)
        _add_comment(label_cell, note_txt)
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        value_parts = _split_overview_value_lines(value_txt)
        for value_part in value_parts:
            lead_txt, rest_txt = _split_overview_prefix(value_part)
            band_fill = copy(row_fill_a if overview_band_idx % 2 == 0 else row_fill_b)
            _fill_row(row_idx, band_fill, narrative_last_col)
            if lead_txt and rest_txt:
                label_cell = ws.cell(row=row_idx, column=1, value=lead_txt)
                label_cell.font = Font(bold=True, size=font_size)
                label_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=narrative_last_col)
                value_cell = ws.cell(row=row_idx, column=2, value=rest_txt)
                value_cell.font = Font(size=font_size)
                value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                value_text_width = sum(
                    float(ws.column_dimensions[get_column_letter(col)].width or 0.0)
                    for col in range(2, narrative_last_col + 1)
                ) or merged_text_width
                label_width = float(ws.column_dimensions["A"].width or 0.0) or 20.0
                label_height = _estimate_wrapped_row_height(
                    lead_txt,
                    label_width,
                    20,
                    14,
                    min_lines=1,
                    max_lines=4,
                )
                value_height = _summary_narrative_row_height(rest_txt, value_text_width)
                ws.row_dimensions[row_idx].height = max(label_height, value_height)
            else:
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=narrative_last_col)
                value_cell = ws.cell(row=row_idx, column=1, value=value_part)
                value_cell.font = Font(size=font_size)
                value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.row_dimensions[row_idx].height = _summary_narrative_row_height(value_part, merged_text_width)
            ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 0, 24.0)
            overview_band_idx += 1
            row_idx += 1

    section_order: List[str] = []
    section_groups: Dict[str, List[Dict[str, Any]]] = {}
    for rec in metric_rows:
        sec = str(rec.get("Section") or "").strip()
        if not sec:
            continue
        section_groups.setdefault(sec, []).append(rec)
        if sec not in section_order:
            section_order.append(sec)

    for sec in section_order:
        row_idx = _write_section_header(row_idx, sec, float(header_size), "5B9BD5", "FFFFFF")
        band_idx = 0
        for rec in section_groups.get(sec, []):
            metric_txt = str(rec.get("Metric") or "").strip()
            if not metric_txt:
                continue
            value = rec.get("Value")
            note_txt = str(rec.get("Note") or "").strip()
            unit_txt = _summary_unit_from_note(note_txt)
            metric_cell = ws.cell(row=row_idx, column=1, value=metric_txt)
            metric_cell.font = Font(size=font_size)
            metric_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            _fill_row(row_idx, copy(row_fill_a if band_idx % 2 == 0 else row_fill_b))
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            if hasattr(value, "year"):
                value_cell.number_format = "yyyy-mm-dd"
            elif isinstance(value, (int, float)) and pd.notna(value):
                if unit_txt == "%":
                    value_cell.number_format = "0.0%"
                elif unit_txt in {"$m", "$/share"}:
                    value_cell.number_format = "#,##0.000"
                elif unit_txt in {"x", "bps"}:
                    value_cell.number_format = "0.000"
                else:
                    value_cell.number_format = "#,##0.000"
            else:
                value_txt = _normalized_value_text(value)
                value_cell.value = value_txt or "N/A"
                value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if not value_cell.alignment or not value_cell.alignment.wrap_text:
                value_cell.alignment = Alignment(horizontal="center", vertical="center")
            unit_cell = ws.cell(row=row_idx, column=3, value=unit_txt or None)
            unit_cell.alignment = Alignment(horizontal="left", vertical="center")
            comment_txt = _summary_note_comment(note_txt)
            if comment_txt:
                _add_comment(metric_cell, comment_txt)
            ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 18, 22)
            band_idx += 1
            row_idx += 1

    append_summary_freshness_sections(
        ws=ws,
        start_row=row_idx,
        source_filing_freshness=(
            source_filing_freshness
            if isinstance(source_filing_freshness, pd.DataFrame)
            else pd.DataFrame()
        ),
        post_quarter_current_effects=(
            post_quarter_current_effects
            if isinstance(post_quarter_current_effects, pd.DataFrame)
            else pd.DataFrame()
        ),
        font_size=font_size,
        header_size=header_size,
    )
