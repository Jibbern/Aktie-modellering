"""Worksheet render adapter for the Valuation Debt Detail section."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, MutableMapping, Optional

import pandas as pd
from openpyxl.styles import Alignment

from .post_quarter_capital_events import apply_pbi_current_debt_overlay


@dataclass(frozen=True)
class ValuationDebtDetailRenderDeps:
    runtime: MutableMapping[str, Any]


@dataclass(frozen=True)
class ValuationDebtDetailRenderResult:
    next_row: int
    row_debt_detail_hdr: int
    tieout_diff_m: Optional[float]
    debt_tieout_guardrail_triggered: bool
    latest_debt_review: bool
    principal_total_m: Optional[float]
    carrying_total_m: Optional[float]
    debt_current_latest_m: Optional[float]
    debt_long_term_latest_m: Optional[float]
    carrying_minus_principal_m: Optional[float]
    near_term_m: Optional[float]


def render_valuation_debt_detail(
    deps: ValuationDebtDetailRenderDeps,
) -> ValuationDebtDetailRenderResult:
    __rt = deps.runtime
    context_globals = dict(__rt.get("context_globals") or {})

    def _rt_get(name: str) -> Any:
        if name in __rt:
            return __rt[name]
        if name in context_globals:
            return context_globals[name]
        return globals().get(name)

    def _coalesce_row_value(row: pd.Series, *names: str, default: Any = None) -> Any:
        for name in names:
            value = row.get(name)
            if value is None:
                continue
            try:
                if pd.isna(value):
                    continue
            except (TypeError, ValueError):
                pass
            if isinstance(value, str) and not value.strip():
                continue
            return value
        return default

    _row_fill = _rt_get("_row_fill")
    _safe_text_value = _rt_get("_safe_text_value")
    _set_cell_comment_local = _rt_get("_set_cell_comment_local")
    _source_backed_debt_tranches_from_slides = _rt_get("_source_backed_debt_tranches_from_slides")
    bold = _rt_get("bold")
    debt_core_map = _rt_get("debt_core_map")
    debt_current_map = _rt_get("debt_current_map")
    debt_maturity = _rt_get("debt_maturity")
    debt_profile = _rt_get("debt_profile")
    debt_tranches_latest = _rt_get("debt_tranches_latest")
    is_pbi_profile = bool(_rt_get("is_pbi_profile"))
    post_quarter_capital_events = _rt_get("post_quarter_capital_events")
    header_fill = _rt_get("header_fill")
    qs = _rt_get("qs")
    r = int(_rt_get("r"))
    section_fill = _rt_get("section_fill")
    slides_debt = _rt_get("slides_debt")
    total_debt_map = _rt_get("total_debt_map")
    ws = _rt_get("ws")

    # Debt detail (latest)
    tieout_diff_m: Optional[float] = None
    row_debt_detail_hdr = r
    ws[f"A{r}"] = "Debt Detail (latest)"
    ws[f"A{r}"].font = bold
    _row_fill(r, section_fill)
    r += 1
    pbi_event: Optional[pd.Series] = None
    if (
        is_pbi_profile
        and isinstance(post_quarter_capital_events, pd.DataFrame)
        and not post_quarter_capital_events.empty
    ):
        pbi_events = post_quarter_capital_events[
            post_quarter_capital_events.get("event_type", pd.Series(dtype=object))
            .astype(str)
            .eq("refinancing_redemption")
        ]
        if not pbi_events.empty:
            pbi_event = pbi_events.iloc[-1]
            ws.cell(
                row=r,
                column=1,
                value="Current / post-quarter principal structure; reported Q1 history unchanged",
            )
            try:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
            except Exception:
                pass
            ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
    debt_header_map = {
        1: "Year/Label",
        2: "Principal due ($m)",
        3: "Rate type",
        4: "Coupon/Spread %",
        6: "Maturity",
        7: "Conversion price",
        9: "Added shares on full conversion (m)",
        12: "Concurrent repurchased shares (m)",
        15: "Basis",
        17: "Source",
    }
    for col_idx, h in debt_header_map.items():
        cell = ws.cell(row=r, column=col_idx, value=h)
        cell.font = bold
        cell.fill = header_fill
    try:
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    except Exception:
        pass
    try:
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
    except Exception:
        pass
    try:
        ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=14)
    except Exception:
        pass
    try:
        ws.merge_cells(start_row=r, start_column=15, end_row=r, end_column=16)
    except Exception:
        pass
    try:
        ws.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)
    except Exception:
        pass
    r += 1
    latest_debt_review = False
    debt_tieout_guardrail_triggered = False
    carrying_total_m: Optional[float] = None
    principal_total_m: Optional[float] = None
    carrying_minus_principal_m: Optional[float] = None
    debt_current_latest_m: Optional[float] = None
    debt_long_term_latest_m: Optional[float] = None
    near_term_m: Optional[float] = None
    if debt_tranches_latest is not None and not debt_tranches_latest.empty:
        df = debt_tranches_latest.copy()
        latest_debt_review = bool(
            ("source_kind" in df.columns and df["source_kind"].astype(str).str.lower().eq("qa_guardrail").any())
            or ("tranche_name" in df.columns and df["tranche_name"].astype(str).str.contains("Needs review", case=False, na=False).any())
        )
        debt_tieout_guardrail_triggered = bool(latest_debt_review)
        source_backed_debt_rows = pd.DataFrame()
        if latest_debt_review:
            source_backed_debt_rows = _source_backed_debt_tranches_from_slides(
                slides_debt,
                qs[-1] if qs else None,
            )
            if not source_backed_debt_rows.empty:
                df = source_backed_debt_rows.copy()
        if pbi_event is not None:
            df = apply_pbi_current_debt_overlay(df, pbi_event)
        sum_listed_principal = 0.0
        near_term = 0.0
        carrying_total_m = None
        principal_total_m = None
        carrying_minus_principal_m = None
        debt_current_latest_m = None
        debt_long_term_latest_m = None
        if debt_profile is not None and not debt_profile.empty:
            dpf = debt_profile.copy()
            if "metric" in dpf.columns and "value" in dpf.columns:
                try:
                    v = pd.to_numeric(dpf.loc[dpf["metric"] == "debt_carrying_total", "value"], errors="coerce").dropna()
                    if not v.empty:
                        carrying_total_m = float(v.iloc[-1]) / 1e6
                except Exception:
                    pass
                try:
                    v = pd.to_numeric(dpf.loc[dpf["metric"] == "debt_principal_total", "value"], errors="coerce").dropna()
                    if not v.empty:
                        principal_total_m = float(v.iloc[-1]) / 1e6
                except Exception:
                    pass
                try:
                    v = pd.to_numeric(
                        dpf.loc[dpf["metric"] == "debt_net_discounts_issuance_costs", "value"],
                        errors="coerce",
                    ).dropna()
                    if not v.empty:
                        carrying_minus_principal_m = float(v.iloc[-1]) / 1e6
                except Exception:
                    pass
                try:
                    v = pd.to_numeric(dpf.loc[dpf["metric"] == "debt_current", "value"], errors="coerce").dropna()
                    if not v.empty:
                        debt_current_latest_m = float(v.iloc[-1]) / 1e6
                except Exception:
                    pass
                try:
                    v = pd.to_numeric(dpf.loc[dpf["metric"] == "debt_long_term", "value"], errors="coerce").dropna()
                    if not v.empty:
                        debt_long_term_latest_m = float(v.iloc[-1]) / 1e6
                except Exception:
                    pass
        if latest_debt_review and source_backed_debt_rows.empty and debt_maturity is not None and not debt_maturity.empty:
            for col_idx, header in debt_header_map.items():
                cell = ws.cell(row=r - 1, column=col_idx, value=header)
                cell.font = bold
                cell.fill = header_fill
            try:
                ws.merge_cells(start_row=r - 1, start_column=7, end_row=r - 1, end_column=8)
                ws.merge_cells(start_row=r - 1, start_column=9, end_row=r - 1, end_column=11)
                ws.merge_cells(start_row=r - 1, start_column=12, end_row=r - 1, end_column=14)
                ws.merge_cells(start_row=r - 1, start_column=15, end_row=r - 1, end_column=16)
                ws.merge_cells(start_row=r - 1, start_column=17, end_row=r - 1, end_column=18)
            except Exception:
                pass
            mat_latest = debt_maturity.copy()
            if "quarter" in mat_latest.columns:
                mat_latest["quarter"] = pd.to_datetime(mat_latest["quarter"], errors="coerce")
                mat_latest = mat_latest[mat_latest["quarter"].dt.to_period("Q") == pd.Timestamp(qs[-1]).to_period("Q")]
            max_rows = 8
            mat_latest = mat_latest.head(max_rows)
            basis_label = "principal_excl_issuance_costs"
            if "source_basis" in mat_latest.columns:
                vals = [str(v).strip() for v in mat_latest["source_basis"].dropna().tolist() if str(v).strip()]
                if vals:
                    basis_label = vals[0]
            for _, row in mat_latest.iterrows():
                amt_total = pd.to_numeric(row.get("amount_total"), errors="coerce")
                maturity_value = _coalesce_row_value(
                    row,
                    "maturity_label",
                    "maturity_year",
                )
                ws.cell(row=r, column=1, value=maturity_value)
                ws.cell(row=r, column=2, value=(float(amt_total) / 1e6) if pd.notna(amt_total) else None).number_format = "#,##0.000"
                ws.cell(row=r, column=3, value=None)
                ws.cell(row=r, column=4, value=None)
                ws.cell(row=r, column=6, value=maturity_value)
                ws.cell(row=r, column=7, value=None)
                ws.cell(row=r, column=9, value=None)
                ws.cell(row=r, column=12, value=None)
                try:
                    ws.merge_cells(start_row=r, start_column=15, end_row=r, end_column=16)
                except Exception:
                    pass
                try:
                    ws.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)
                except Exception:
                    pass
                ws.cell(row=r, column=15, value=basis_label)
                ws.cell(
                    row=r,
                    column=17,
                    value=_coalesce_row_value(
                        row,
                        "source_kind",
                        default="Debt_Maturity_Ladder",
                    ),
                )
                if pd.notna(amt_total):
                    sum_listed_principal += float(amt_total)
                r += 1
            r += 1
            ws.cell(
                row=r,
                column=1,
                value="Scheduled repayments / principal basis (excludes debt issuance costs). Carrying debt_core may differ from principal total.",
            )
            try:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
            except Exception:
                pass
            ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            r += 2
        else:
            max_rows = 14
            if len(df) > max_rows:
                df = df.head(max_rows)
            for _, row in df.iterrows():
                ws.cell(row=r, column=1, value=row.get("tranche_name"))
                amt_pr = pd.to_numeric(row.get("amount_principal", row.get("amount")), errors="coerce")
                amt_ca = pd.to_numeric(row.get("amount_carrying"), errors="coerce")
                if pd.notna(amt_pr):
                    sum_listed_principal += float(amt_pr)
                    if bool(row.get("near_term")):
                        near_term += float(amt_pr)
                ws.cell(row=r, column=2, value=(float(amt_pr) / 1e6) if pd.notna(amt_pr) else None).number_format = "#,##0.000"
                m_disp = _coalesce_row_value(row, "maturity_display")
                if m_disp is None and pd.notna(row.get("maturity_year")):
                    m_disp = str(int(pd.to_numeric(row.get("maturity_year"), errors="coerce")))
                ws.cell(
                    row=r,
                    column=3,
                    value=_coalesce_row_value(row, "rate_type", "instrument_type"),
                )
                coupon_or_spread = pd.to_numeric(
                    _coalesce_row_value(row, "coupon_pct", "spread_pct"),
                    errors="coerce",
                )
                ws.cell(
                    row=r,
                    column=4,
                    value=float(coupon_or_spread) if pd.notna(coupon_or_spread) else None,
                )
                ws.cell(row=r, column=6, value=m_disp)
                conv_price = pd.to_numeric(row.get("conversion_price"), errors="coerce")
                conv_shares = pd.to_numeric(row.get("shares_on_full_conversion"), errors="coerce")
                rep_shares = pd.to_numeric(row.get("concurrent_repurchase_shares"), errors="coerce")
                ws.cell(row=r, column=7, value=float(conv_price) if pd.notna(conv_price) else None).number_format = "$#,##0.00"
                ws.cell(row=r, column=9, value=(float(conv_shares) / 1e6) if pd.notna(conv_shares) else None).number_format = "#,##0.000"
                ws.cell(row=r, column=12, value=(float(rep_shares) / 1e6) if pd.notna(rep_shares) else None).number_format = "#,##0.000"
                debt_dilution_note = _safe_text_value(row.get("dilution_structure_note"))
                if debt_dilution_note:
                    try:
                        note_cell = ws.cell(row=r, column=9)
                        if not pd.notna(conv_shares):
                            note_cell = ws.cell(row=r, column=1)
                        _set_cell_comment_local(note_cell, debt_dilution_note)
                    except Exception:
                        pass
                try:
                    ws.merge_cells(start_row=r, start_column=15, end_row=r, end_column=16)
                except Exception:
                    pass
                try:
                    ws.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)
                except Exception:
                    pass
                basis_display = _coalesce_row_value(
                    row,
                    "source_basis",
                    default="principal_tranche_sum",
                )
                if "within 24 months of latest quarter end" in str(basis_display):
                    basis_display = "source-backed"
                ws.cell(row=r, column=15, value=basis_display)
                ws.cell(
                    row=r,
                    column=17,
                    value=_coalesce_row_value(
                        row,
                        "source_kind",
                        default="Debt_Tranches_Latest",
                    ),
                )
                row_conv_note = _safe_text_value(row.get("conversion_terms_note"))
                if (pd.notna(conv_price) or pd.notna(conv_shares)) and row_conv_note:
                    try:
                        comment_parts = [row_conv_note]
                        if pd.notna(pd.to_numeric(row.get("concurrent_repurchase_amount"), errors="coerce")):
                            comment_parts.append(
                                f"Concurrent repurchase: ${float(pd.to_numeric(row.get('concurrent_repurchase_amount'), errors='coerce'))/1e6:,.1f}m"
                            )
                        if pd.notna(pd.to_numeric(row.get("concurrent_repurchase_shares"), errors="coerce")):
                            comment_parts.append(
                                f"Concurrent repurchase shares: {float(pd.to_numeric(row.get('concurrent_repurchase_shares'), errors='coerce'))/1e6:,.1f}m"
                            )
                        if _safe_text_value(row.get("hedge_or_call_spread")):
                            comment_parts.append(f"Hedge/call spread: {_safe_text_value(row.get('hedge_or_call_spread'))}")
                        if _safe_text_value(row.get("settlement_type")):
                            comment_parts.append(f"Settlement: {_safe_text_value(row.get('settlement_type'))}")
                        if _safe_text_value(row.get("conversion_conditions_note")):
                            comment_parts.append(f"Conditions: {_safe_text_value(row.get('conversion_conditions_note'))}")
                        _set_cell_comment_local(
                            ws.cell(row=r, column=17),
                            "\n".join([x for x in comment_parts if x])
                            + f"\n\nSource: {_safe_text_value(row.get('conversion_terms_source')) or _safe_text_value(row.get('source_kind'))}",
                        )
                    except Exception:
                        pass
                r += 1
        # visual spacer before tie-out rows
        r += 1
        # principal/carrying reconciliation rows
        if pbi_event is not None:
            principal_total_m = sum_listed_principal / 1e6 if sum_listed_principal else None
        else:
            principal_total_m = principal_total_m if principal_total_m is not None else (sum_listed_principal / 1e6 if sum_listed_principal else None)
        if carrying_total_m is None:
            debt_core_latest = debt_core_map.get(pd.Timestamp(qs[-1])) if qs else None
            carrying_total_m = (float(debt_core_latest) / 1e6) if debt_core_latest is not None else None
        debt_current_latest = None
        if debt_current_latest_m is None:
            debt_current_latest = debt_current_map.get(pd.Timestamp(qs[-1])) if qs else None
            debt_current_latest_m = (float(debt_current_latest) / 1e6) if debt_current_latest is not None else None
        if carrying_total_m is not None and debt_current_latest_m is not None:
            # Display a non-duplicative long-term carrying view.  Some
            # source schedules expose total/core carrying debt separately
            # from current debt; showing that total as "long-term" makes
            # the section look like current + long-term double counts.
            debt_long_term_latest_m = max(0.0, float(carrying_total_m) - float(debt_current_latest_m))
        elif debt_long_term_latest_m is None:
            total_debt_latest = total_debt_map.get(pd.Timestamp(qs[-1])) if qs else None
            if total_debt_latest is not None:
                debt_st_latest = debt_current_latest if debt_current_latest is not None else 0.0
                debt_long_term_latest_m = (float(total_debt_latest) - float(debt_st_latest)) / 1e6
        if carrying_minus_principal_m is None and carrying_total_m is not None and principal_total_m is not None:
            carrying_minus_principal_m = float(carrying_total_m) - float(principal_total_m)

        ws.cell(row=r, column=1, value="Principal total ($m)").font = bold
        ws.cell(row=r, column=2, value=principal_total_m).number_format = "#,##0.000"
        r += 1
        carrying_label = "Reported Q1 carrying debt_core ($m)" if pbi_event is not None else "Carrying debt_core ($m)"
        ws.cell(row=r, column=1, value=carrying_label).font = bold
        ws.cell(row=r, column=2, value=carrying_total_m).number_format = "#,##0.000"
        r += 1
        debt_current_label = "Reported Q1 debt current ($m)" if pbi_event is not None else "Debt current ($m)"
        ws.cell(row=r, column=1, value=debt_current_label).font = bold
        ws.cell(row=r, column=2, value=debt_current_latest_m).number_format = "#,##0.000"
        r += 1
        debt_long_term_label = (
            "Reported Q1 debt long-term carrying (core-current, $m)"
            if pbi_event is not None
            else "Debt long-term carrying (core-current, $m)"
        )
        ws.cell(row=r, column=1, value=debt_long_term_label).font = bold
        ws.cell(row=r, column=2, value=debt_long_term_latest_m).number_format = "#,##0.000"
        r += 1
        ws.cell(row=r, column=1, value="Carrying less principal ($m)").font = bold
        ws.cell(row=r, column=2, value=carrying_minus_principal_m).number_format = "#,##0.000"
        tieout_diff_m = carrying_minus_principal_m
        r += 1
        if carrying_total_m not in (None, 0) and carrying_minus_principal_m is not None:
            ws.cell(row=r, column=1, value="Carrying less principal (%)").font = bold
            ws.cell(row=r, column=2, value=(float(carrying_minus_principal_m) / float(carrying_total_m))).number_format = "0.0%"
            r += 1
        ws.cell(
            row=r,
            column=1,
            value="Near-term maturities (within 24m of latest quarter end, principal, $m)",
        ).font = bold
        near_term_m = near_term / 1e6
        ws.cell(row=r, column=2, value=near_term_m).number_format = "#,##0.000"
        r += 1
        if debt_tieout_guardrail_triggered:
            warn_txt = (
                "WARN: tranche tie-out guardrail fired; source-backed principal schedule is shown for audit, "
                "while carrying debt_core remains the valuation debt basis."
            )
            ws.cell(row=r, column=1, value=warn_txt)
            try:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
            except Exception:
                pass
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            r += 1
        if pbi_event is not None:
            panel_row = row_debt_detail_hdr
            panel_label_start = 19
            panel_label_end = 22
            panel_value_start = 23
            panel_value_end = 26
            try:
                ws.merge_cells(
                    start_row=panel_row,
                    start_column=panel_label_start,
                    end_row=panel_row,
                    end_column=panel_value_end,
                )
            except Exception:
                pass
            ws.cell(
                row=panel_row,
                column=panel_label_start,
                value="Post-quarter refinancing overlay / not in reported Q1 values",
            ).font = bold
            for column in range(panel_label_start, panel_value_end + 1):
                ws.cell(row=panel_row, column=column).fill = section_fill
            source_note = (
                f"Source: {pbi_event.get('filing_type')} accession {pbi_event.get('accession')}\n"
                f"{pbi_event.get('source_paths')}"
            )
            try:
                _set_cell_comment_local(
                    ws.cell(row=panel_row, column=panel_label_start),
                    source_note,
                )
            except Exception:
                pass
            panel_row += 1
            event_rows = (
                ("2027 Senior Notes redeemed ($m)", -float(pbi_event["principal_redeemed"]) / 1e6),
                ("Incremental Term Loan A ($m)", float(pbi_event["incremental_term_loan"]) / 1e6),
                ("Term Loan A total after amendment ($m)", float(pbi_event["term_loan_total"]) / 1e6),
                ("Gross principal debt delta before fees/costs/other sources ($m)", float(pbi_event["gross_principal_delta"]) / 1e6),
                ("Cash / current net debt", "Unresolved / manual review"),
                ("Automatic pro-forma net debt adjustment", "Disabled / manual"),
                ("Next scheduled maturity", pbi_event["next_scheduled_maturity"]),
                ("Term Loan A maturity", "May 18, 2031"),
            )
            for label, value in event_rows:
                try:
                    ws.merge_cells(
                        start_row=panel_row,
                        start_column=panel_label_start,
                        end_row=panel_row,
                        end_column=panel_label_end,
                    )
                    ws.merge_cells(
                        start_row=panel_row,
                        start_column=panel_value_start,
                        end_row=panel_row,
                        end_column=panel_value_end,
                    )
                except Exception:
                    pass
                ws.cell(row=panel_row, column=panel_label_start, value=label)
                ws.cell(row=panel_row, column=panel_value_start, value=value)
                ws.cell(
                    row=panel_row,
                    column=panel_label_start,
                ).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(
                    row=panel_row,
                    column=panel_value_start,
                ).alignment = Alignment(wrap_text=True, vertical="top")
                if isinstance(value, (int, float)):
                    ws.cell(
                        row=panel_row,
                        column=panel_value_start,
                    ).number_format = "#,##0.000"
                panel_row += 1

    return ValuationDebtDetailRenderResult(
        next_row=r,
        row_debt_detail_hdr=row_debt_detail_hdr,
        tieout_diff_m=tieout_diff_m,
        debt_tieout_guardrail_triggered=debt_tieout_guardrail_triggered,
        latest_debt_review=latest_debt_review,
        principal_total_m=principal_total_m,
        carrying_total_m=carrying_total_m,
        debt_current_latest_m=debt_current_latest_m,
        debt_long_term_latest_m=debt_long_term_latest_m,
        carrying_minus_principal_m=carrying_minus_principal_m,
        near_term_m=near_term_m,
    )
