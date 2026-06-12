"""Final cross-panel worksheet layout pass for the Valuation sheet."""
from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from typing import Any, Dict, MutableMapping, Optional

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


@dataclass(frozen=True)
class ValuationFinalLayoutDeps:
    runtime: MutableMapping[str, Any]


@dataclass(frozen=True)
class ValuationFinalLayoutResult:
    freeze_panes: str
    valuation_start_row: int
    max_row: int
    max_column: int


def apply_valuation_final_layout(
    deps: ValuationFinalLayoutDeps,
) -> ValuationFinalLayoutResult:
    __rt = deps.runtime
    context_globals = dict(__rt.get("context_globals") or {})

    def _rt_get(name: str) -> Any:
        if name in __rt:
            return __rt[name]
        if name in context_globals:
            return context_globals[name]
        return globals().get(name)

    _anf_clean_visible_ui_text = _rt_get("_anf_clean_visible_ui_text")
    _anf_clear_valuation_side_panels = _rt_get("_anf_clear_valuation_side_panels")
    _px_to_width = _rt_get("_px_to_width")
    _record_writer_elapsed = _rt_get("_record_writer_elapsed")
    _record_writer_substage = _rt_get("_record_writer_substage")
    _set_cell_comment = _rt_get("_set_cell_comment")
    _updated_font = _rt_get("_updated_font")
    _write_anf_valuation_side_panel = _rt_get("_write_anf_valuation_side_panel")
    actuals_row = _rt_get("actuals_row")
    additive_panel_end = _rt_get("additive_panel_end")
    bold = _rt_get("bold")
    col_exact_start = _rt_get("col_exact_start")
    col_guidance_start = _rt_get("col_guidance_start")
    col_horizon_start = _rt_get("col_horizon_start")
    col_metric_start = _rt_get("col_metric_start")
    col_stated_start = _rt_get("col_stated_start")
    convert_header_end_col = _rt_get("convert_header_end_col")
    data_start_row = _rt_get("data_start_row")
    dcf_interp_col = _rt_get("dcf_interp_col")
    dcf_label_col = _rt_get("dcf_label_col")
    dcf_value_col = _rt_get("dcf_value_col")
    driver_value_col = _rt_get("driver_value_col")
    font_size = _rt_get("font_size")
    grid_col_start = _rt_get("grid_col_start")
    grid_layout_width = _rt_get("grid_layout_width")
    grid_start = _rt_get("grid_start")
    guidance_snapshot_header_rows = _rt_get("guidance_snapshot_header_rows")
    header_fill = _rt_get("header_fill")
    header_size = _rt_get("header_size")
    hv_panel_label_col = _rt_get("hv_panel_label_col")
    hv_panel_val_col = _rt_get("hv_panel_val_col")
    input_hint_col = _rt_get("input_hint_col")
    input_label_col = _rt_get("input_label_col")
    input_value_col = _rt_get("input_value_col")
    is_anf_profile = _rt_get("is_anf_profile")
    last_col = _rt_get("last_col")
    last_col_letter = _rt_get("last_col_letter")
    market_interp_col = _rt_get("market_interp_col")
    market_label_col = _rt_get("market_label_col")
    market_value_col = _rt_get("market_value_col")
    output_interp_col = _rt_get("output_interp_col")
    output_label_col = _rt_get("output_label_col")
    output_value_col = _rt_get("output_value_col")
    panel_col = _rt_get("panel_col")
    panel_col_end = _rt_get("panel_col_end")
    panel_col_start = _rt_get("panel_col_start")
    panel_row_start = _rt_get("panel_row_start")
    qadj_text_col = _rt_get("qadj_text_col")
    quarter_row = _rt_get("quarter_row")
    row_asof = _rt_get("row_asof")
    row_convert_hdr = _rt_get("row_convert_hdr")
    row_dcf_end = _rt_get("row_dcf_end")
    row_dcf_eq = _rt_get("row_dcf_eq")
    row_dcf_hdr = _rt_get("row_dcf_hdr")
    row_dcf_sens_hdr = _rt_get("row_dcf_sens_hdr")
    row_dcf_sens_last_row = _rt_get("row_dcf_sens_last_row")
    row_debt_detail_hdr = _rt_get("row_debt_detail_hdr")
    row_drv_fcf = _rt_get("row_drv_fcf")
    row_drv_hdr = _rt_get("row_drv_hdr")
    row_drv_lev = _rt_get("row_drv_lev")
    row_drv_margin = _rt_get("row_drv_margin")
    row_drv_rev = _rt_get("row_drv_rev")
    row_fill_elapsed = _rt_get("row_fill_elapsed")
    row_flags_hdr = _rt_get("row_flags_hdr")
    row_hv_cap_hdr_dyn = _rt_get("row_hv_cap_hdr_dyn")
    row_hv_hdr_dyn = _rt_get("row_hv_hdr_dyn")
    row_hv_obs_hdr_dyn = _rt_get("row_hv_obs_hdr_dyn")
    row_market_hdr = _rt_get("row_market_hdr")
    row_mi_hdr = _rt_get("row_mi_hdr")
    row_mi_tbl_hdr = _rt_get("row_mi_tbl_hdr")
    row_mi_toggle = _rt_get("row_mi_toggle")
    row_mktcap = _rt_get("row_mktcap")
    row_operating_hdr = _rt_get("row_operating_hdr")
    row_ptbv = _rt_get("row_ptbv")
    row_qadj_ev = _rt_get("row_qadj_ev")
    row_qadj_ev_adj = _rt_get("row_qadj_ev_adj")
    row_qadj_hdr = _rt_get("row_qadj_hdr")
    row_qadj_yield = _rt_get("row_qadj_yield")
    row_req_adj_ebitda = _rt_get("row_req_adj_ebitda")
    row_req_owner_delta = _rt_get("row_req_owner_delta")
    row_scn_eq_fcf = _rt_get("row_scn_eq_fcf")
    row_scn_hdr = _rt_get("row_scn_hdr")
    row_scn_profile = _rt_get("row_scn_profile")
    row_thesis_end = _rt_get("row_thesis_end")
    row_thesis_hdr = _rt_get("row_thesis_hdr")
    row_toggle_hdr = _rt_get("row_toggle_hdr")
    row_trend_hdr = _rt_get("row_trend_hdr")
    row_write_elapsed = _rt_get("row_write_elapsed")
    scn_interp_col = _rt_get("scn_interp_col")
    scn_label_col = _rt_get("scn_label_col")
    scn_value_col = _rt_get("scn_value_col")
    thick = _rt_get("thick")
    title_fill = _rt_get("title_fill")
    valuation_header_row = _rt_get("valuation_header_row")
    valuation_inputs_row = _rt_get("valuation_inputs_row")
    valuation_render_started = _rt_get("valuation_render_started")
    visible_hv_flags_hdr_row = _rt_get("visible_hv_flags_hdr_row")
    wb = _rt_get("wb")
    ws = _rt_get("ws")

    def _apply_valuation_layout(ws_local: Any) -> None:
        # Keep the quarter table D:M untouched; B:C are explicit across tickers.
        protected_cols = {get_column_letter(c) for c in range(4, 14)}  # D..M

        def _set_layout_width(col_letter: str, width: float) -> None:
            if col_letter in protected_cols:
                return
            ws_local.column_dimensions[col_letter].width = width

        _set_layout_width("A", 50.0)
        ws_local.column_dimensions["B"].width = _px_to_width(92)
        ws_local.column_dimensions["C"].width = _px_to_width(92)
        _set_layout_width("N", _px_to_width(91))
        # Match the ANF side-panel column system so PBI/GPRE render as the
        # same platform while keeping company-specific content.
        for col_letter, width in {
            "O": 18.0,
            "P": 16.0,
            "Q": 14.0,
            "R": 16.0,
            "S": 16.0,
            "T": 16.0,
            "U": 16.0,
            "V": 16.0,
            "W": 16.0,
            "X": 16.0,
            "Y": 16.0,
            "Z": 16.0,
            "AA": 18.0,
            "AB": 18.0,
            "AC": 18.0,
        }.items():
            _set_layout_width(col_letter, width)

        # Find valuation start dynamically from column B.
        valuation_start_row = valuation_header_row
        for rr in range(1, min(ws_local.max_row, 400) + 1):
            if str(ws_local.cell(row=rr, column=2).value or "").strip().lower() == "valuation":
                valuation_start_row = rr
                break

        # Guidance block readability (rows above valuation panel in O:AB).
        for rr in range(1, valuation_start_row):
            if row_operating_hdr and row_operating_hdr <= rr <= row_thesis_end:
                continue
            has_guidance = any(ws_local.cell(row=rr, column=cc).value not in (None, "") for cc in range(panel_col_start, panel_col_end + 1))
            if not has_guidance:
                continue
            metric_cell = ws_local.cell(row=rr, column=col_metric_start)
            stated_cell = ws_local.cell(row=rr, column=col_stated_start)
            horizon_cell = ws_local.cell(row=rr, column=col_horizon_start)
            value_cell = ws_local.cell(row=rr, column=col_guidance_start)
            exact_cell = ws_local.cell(row=rr, column=col_exact_start)

            metric_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            stated_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            horizon_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            exact_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


        # As-of display guard (keep date in the valuation input cell, no address moves).
        asof_cell = ws_local.cell(row=row_asof, column=input_value_col)
        asof_cell.number_format = "yyyy-mm-dd"
        asof_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True)
        try:
            asof_dt = pd.to_datetime(asof_cell.value, errors="coerce")
            if pd.notna(asof_dt):
                _set_cell_comment(asof_cell, f"As of: {pd.Timestamp(asof_dt).strftime('%Y-%m-%d')}")
        except Exception:
            pass

        # Build protected named-range cells to avoid risky merges.
        valuation_named_cells: set[tuple[int, int]] = set()
        try:
            try:
                _defined_names = list(wb.defined_names.values())
            except Exception:
                _defined_names = list(getattr(wb.defined_names, "definedName", []) or [])
            for dn in _defined_names:
                try:
                    destinations = list(dn.destinations)
                except Exception:
                    continue
                for sht, coord in destinations:
                    try:
                        sht_name = str(sht).strip("'")
                    except Exception:
                        continue
                    if sht_name != ws_local.title:
                        continue
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(str(coord))
                    except Exception:
                        continue
                    if (max_row - min_row) > 250 or (max_col - min_col) > 60:
                        continue
                    for rr in range(min_row, max_row + 1):
                        for cc in range(min_col, max_col + 1):
                            valuation_named_cells.add((rr, cc))
        except Exception:
            valuation_named_cells = set()

        protected_block_cols: set[int] = set()
        for cc in range(input_label_col, input_value_col + 1):
            protected_block_cols.add(cc)
        for cc in [output_value_col]:
            protected_block_cols.add(cc)
        for cc in range(grid_col_start, grid_col_start + grid_layout_width):
            protected_block_cols.add(cc)

        def _cell_in_merged_range(row_idx: int, col_idx: int) -> bool:
            for mr in ws_local.merged_cells.ranges:
                if mr.min_row <= row_idx <= mr.max_row and mr.min_col <= col_idx <= mr.max_col:
                    return True
            return False

        def _safe_unmerge_from_start(row_idx: int, col_idx: int) -> None:
            for mr in list(ws_local.merged_cells.ranges):
                if mr.min_row == row_idx and mr.max_row == row_idx and mr.min_col == col_idx:
                    try:
                        ws_local.unmerge_cells(str(mr))
                    except Exception:
                        pass

        def _is_safe_blank(row_idx: int, col_idx: int) -> bool:
            c = ws_local.cell(row=row_idx, column=col_idx)
            if c.value is not None:
                return False
            if getattr(c, "data_type", "") == "f":
                return False
            if c.comment is not None:
                return False
            if (row_idx, col_idx) in valuation_named_cells:
                return False
            if _cell_in_merged_range(row_idx, col_idx):
                return False
            if col_idx in protected_block_cols:
                return False
            return True

        def _merge_text_cell_horiz(
            row_idx: int,
            start_col: int,
            max_end_col: int,
            max_span: int = 6,
        ) -> int:
            end_col = start_col
            cap_col = min(max_end_col, start_col + max_span)
            for cc in range(start_col + 1, cap_col + 1):
                if not _is_safe_blank(row_idx, cc):
                    break
                end_col = cc
            if end_col > start_col:
                _safe_unmerge_from_start(row_idx, start_col)
                try:
                    ws_local.merge_cells(
                        start_row=row_idx,
                        start_column=start_col,
                        end_row=row_idx,
                        end_column=end_col,
                    )
                except Exception:
                    end_col = start_col
            return end_col

        def _merged_span_end_col(row_idx: int, start_col: int) -> int:
            for mr in ws_local.merged_cells.ranges:
                if mr.min_row == row_idx and mr.max_row == row_idx and mr.min_col == start_col:
                    return int(mr.max_col)
            return start_col

        def _span_width_chars(row_idx: int, start_col: int) -> float:
            end_col = _merged_span_end_col(row_idx, start_col)
            total = 0.0
            for cc in range(start_col, end_col + 1):
                w = ws_local.column_dimensions[get_column_letter(cc)].width
                total += float(w) if w is not None else 10.0
            return max(10.0, total)

        def _text_line_count(text_val: Any, width_chars: float) -> int:
            txt = str(text_val or "")
            if not txt:
                return 0
            if txt.startswith("="):
                return 1
            txt = re.sub(r"\s+", " ", txt).strip()
            if not txt:
                return 0
            chars_per_line = max(20, int(width_chars * 1.10) - 4)
            return max(1, int((len(txt) + chars_per_line - 1) / chars_per_line)) + txt.count("\n")

        def _line_height(lines: int) -> float:
            if lines <= 1:
                return 18.0
            if lines == 2:
                return 28.0
            if lines == 3:
                return 40.0
            return 52.0

        def _force_merge_row(row_idx: int, start_col: int, end_col: int) -> None:
            if end_col <= start_col:
                return
            _safe_unmerge_from_start(row_idx, start_col)
            for mr in list(ws_local.merged_cells.ranges):
                try:
                    same_range = (
                        mr.min_row == row_idx
                        and mr.max_row == row_idx
                        and mr.min_col == start_col
                        and mr.max_col == end_col
                    )
                    overlaps_other = (
                        mr.min_row <= row_idx <= mr.max_row
                        and mr.min_col <= end_col
                        and mr.max_col >= start_col
                    )
                    if overlaps_other and not same_range:
                        return
                except Exception:
                    continue
            try:
                ws_local.merge_cells(
                    start_row=row_idx,
                    start_column=start_col,
                    end_row=row_idx,
                    end_column=end_col,
                )
            except Exception:
                pass

        # 3) Text fields in x-direction using safe merges.
        text_rows_min = valuation_start_row + 1
        text_rows_max = max(row_dcf_end, row_qadj_yield)
        row_target_heights: Dict[int, float] = {}

        for rr in range(panel_row_start, valuation_start_row):
            if row_operating_hdr and row_operating_hdr <= rr <= row_thesis_end:
                continue
            if not any(ws_local.cell(row=rr, column=cc).value not in (None, "") for cc in range(panel_col_start, panel_col_end + 1)):
                continue
            # The top guidance snapshot is intentionally compact, but do
            # not shrink rows that a ticker-specific side panel already
            # made taller to keep the shared quarterly grid readable.
            current_height = ws_local.row_dimensions[rr].height
            if current_height is None or float(current_height) < 18.0:
                ws_local.row_dimensions[rr].height = 18.0

        # Header/title spans (fills should cover full visible area).
        _force_merge_row(valuation_inputs_row, input_label_col, input_value_col - 1)      # Inputs B:C
        _force_merge_row(valuation_inputs_row, output_label_col, output_value_col - 1)    # Outputs K:M
        _force_merge_row(valuation_header_row, input_label_col, 19)                       # Valuation B:S
        _force_merge_row(valuation_inputs_row, output_interp_col, 19)                     # Interpretation O:S
        _force_merge_row(row_market_hdr, market_label_col, 21)                            # What Market Is Pricing L:U
        _force_merge_row(row_scn_hdr, scn_label_col, scn_interp_col + 5)                  # Trigger Scenarios B:K
        _force_merge_row(row_dcf_hdr, dcf_label_col, dcf_value_col)                       # DCF (optional module) G:J
        _force_merge_row(grid_start, 2, 5)                                                # Valuation Sensitivity Grid B:E
        _force_merge_row(row_dcf_sens_hdr, dcf_label_col, dcf_label_col + 5)             # DCF Sensitivity G:L
        _force_merge_row(row_drv_hdr, 2, 11)                                              # Scenario drivers B:K
        _force_merge_row(row_toggle_hdr, 2, 5)                                            # Quality toggles B:E
        _force_merge_row(row_qadj_hdr, 2, 11)                                             # Q-adjusted B:K
        _force_merge_row(row_mi_hdr, 16, 19)                                              # Market-implied P:S
        _force_merge_row(row_mi_tbl_hdr, 17, 18)                                          # Implied g Q:R

        def _apply_major_section_band(row_idx: int, start_col: int, end_col: int) -> None:
            if row_idx <= 0 or end_col < start_col:
                return
            for cc in range(start_col, end_col + 1):
                cell = ws_local.cell(row=row_idx, column=cc)
                cell.fill = copy(title_fill)
                if cell.value not in (None, ""):
                    cell.font = Font(
                        bold=True,
                        size=float(getattr(cell.font, "size", header_size) or header_size),
                        color="FFFFFF",
                    )

        _apply_major_section_band(3, 1, 13)
        _apply_major_section_band(visible_hv_flags_hdr_row, 1, 13)
        _apply_major_section_band(row_hv_obs_hdr_dyn, 1, 13)
        _apply_major_section_band(row_hv_cap_hdr_dyn, 1, 13)
        _apply_major_section_band(row_hv_hdr_dyn, hv_panel_label_col, hv_panel_val_col)
        _apply_major_section_band(row_debt_detail_hdr, 1, 18)
        _apply_major_section_band(valuation_header_row, input_label_col, 19)
        _apply_major_section_band(row_convert_hdr, 12, convert_header_end_col)
        _apply_major_section_band(row_operating_hdr, panel_col_start, additive_panel_end)
        _apply_major_section_band(row_thesis_hdr, panel_col_start, additive_panel_end)
        _apply_major_section_band(row_trend_hdr, panel_col, panel_col + 3)
        _apply_major_section_band(row_flags_hdr, panel_col, panel_col + 8)
        for rr in guidance_snapshot_header_rows:
            _apply_major_section_band(rr, panel_col_start, panel_col_end)
        # Explicit value headers for numeric columns in the top valuation row.
        for value_col in (input_value_col, output_value_col):
            for mrange in list(ws_local.merged_cells.ranges):
                if mrange.min_row <= valuation_inputs_row <= mrange.max_row and mrange.min_col <= value_col <= mrange.max_col:
                    try:
                        ws_local.unmerge_cells(str(mrange))
                    except Exception:
                        pass
            v_head = ws_local.cell(row=valuation_inputs_row, column=value_col, value="Value")
            v_head.font = bold
            v_head.fill = header_fill
            v_head.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

        for rr in range(text_rows_min, text_rows_max + 1):
            h_cell = ws_local.cell(row=rr, column=input_hint_col)
            if h_cell.value not in (None, ""):
                # Hint: expand over free space up to block divider (typically F:G).
                _merge_text_cell_horiz(
                    rr,
                    input_hint_col,
                    max(input_hint_col, output_label_col - 1),
                    max_span=max(1, output_label_col - input_hint_col),
                )
                h_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                h_lines = _text_line_count(h_cell.value, _span_width_chars(rr, input_hint_col))
                row_target_heights[rr] = max(row_target_heights.get(rr, 0.0), _line_height(h_lines))

            # Outputs labels: I:J (based on configured output columns)
            out_label = ws_local.cell(row=rr, column=output_label_col)
            if rr >= row_mktcap and rr <= row_ptbv and out_label.value not in (None, ""):
                _merge_text_cell_horiz(
                    rr,
                    output_label_col,
                    max(output_label_col, output_value_col - 1),
                    max_span=max(1, output_value_col - output_label_col),
                )
                out_label.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            out_interp = ws_local.cell(row=rr, column=output_interp_col)
            if rr >= row_mktcap and rr <= row_ptbv and out_interp.value not in (None, ""):
                _force_merge_row(rr, output_interp_col, 19)  # O:S
                out_interp.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                out_lines = _text_line_count(out_interp.value, _span_width_chars(rr, output_interp_col))
                row_target_heights[rr] = max(row_target_heights.get(rr, 0.0), _line_height(out_lines))

            # What market labels: L:N with values in O.
            mkt_label = ws_local.cell(row=rr, column=market_label_col)
            if mkt_label.value not in (None, ""):
                if rr >= row_market_hdr and rr <= row_req_owner_delta:
                    _merge_text_cell_horiz(
                        rr,
                        market_label_col,
                        market_value_col - 1,
                        max_span=max(1, market_value_col - market_label_col - 1),
                    )  # L:O
                mkt_label.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Market interpretation block: Q:U for market requirement rows.
            mkt_interp = ws_local.cell(row=rr, column=market_interp_col)  # Q
            if rr >= row_req_adj_ebitda and rr <= row_req_owner_delta and mkt_interp.value not in (None, ""):
                _merge_text_cell_horiz(
                    rr,
                    market_interp_col,
                    market_interp_col + 4,  # U
                    max_span=4,
                )
                mkt_interp.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                mkt_lines = _text_line_count(mkt_interp.value, _span_width_chars(rr, market_interp_col))
                row_target_heights[rr] = max(row_target_heights.get(rr, 0.0), _line_height(mkt_lines))

            # Trigger Scenarios labels P:R and interpretation T:Y.
            scn_label = ws_local.cell(row=rr, column=scn_label_col)
            if rr >= row_scn_hdr and rr <= row_scn_eq_fcf and scn_label.value not in (None, ""):
                _merge_text_cell_horiz(rr, scn_label_col, scn_value_col - 1, max_span=3)  # P:R
                scn_label.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            scn_interp = ws_local.cell(row=rr, column=scn_interp_col)  # F
            if rr >= row_scn_profile and rr <= row_scn_eq_fcf and scn_interp.value not in (None, ""):
                _force_merge_row(rr, scn_interp_col, scn_interp_col + 5)  # F:K
                scn_interp.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                scn_lines = _text_line_count(scn_interp.value, _span_width_chars(rr, scn_interp_col))
                row_target_heights[rr] = max(row_target_heights.get(rr, 0.0), _line_height(scn_lines))

        # Scenario drivers values D:K.
        for rr in [row_drv_rev, row_drv_margin, row_drv_fcf, row_drv_lev]:
            dval = ws_local.cell(row=rr, column=driver_value_col)
            if dval.value not in (None, ""):
                _safe_unmerge_from_start(rr, driver_value_col)
                try:
                    can_merge = True
                    for cc in range(driver_value_col + 1, 12):  # E:K
                        cc_cell = ws_local.cell(row=rr, column=cc)
                        if cc_cell.value is not None or getattr(cc_cell, "data_type", "") == "f" or cc_cell.comment is not None:
                            can_merge = False
                            break
                    if can_merge:
                        ws_local.merge_cells(
                            start_row=rr,
                            start_column=driver_value_col,
                            end_row=rr,
                            end_column=11,
                        )
                except Exception:
                    pass
                dval.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                d_lines = _text_line_count(dval.value, _span_width_chars(rr, driver_value_col))
                row_target_heights[rr] = max(row_target_heights.get(rr, 0.0), _line_height(d_lines))

        # Quality-adjusted interpretation text E:K.
        for rr in [row_qadj_ev_adj, row_qadj_ev, row_qadj_yield]:
            qtxt = ws_local.cell(row=rr, column=qadj_text_col)
            if qtxt.value not in (None, ""):
                _safe_unmerge_from_start(rr, qadj_text_col)
                try:
                    ws_local.merge_cells(
                        start_row=rr,
                        start_column=qadj_text_col,
                        end_row=rr,
                        end_column=11,
                    )  # F:K
                except Exception:
                    pass
                qtxt.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                q_lines = _text_line_count(qtxt.value, _span_width_chars(rr, qadj_text_col))
                row_target_heights[rr] = max(row_target_heights.get(rr, 0.0), _line_height(q_lines))

        # DCF interpretation rows (same behavior, no cross-block merge risk).
        # Keep DCF header compact/readable: label spans G:J and interpretation spans K:O.
        try:
            _safe_unmerge_from_start(row_dcf_hdr, dcf_label_col)
            ws_local.merge_cells(
                start_row=row_dcf_hdr,
                start_column=dcf_label_col,
                end_row=row_dcf_hdr,
                end_column=dcf_value_col,
            )
        except Exception:
            pass
        try:
            _safe_unmerge_from_start(row_dcf_hdr, dcf_interp_col)
            ws_local.merge_cells(
                start_row=row_dcf_hdr,
                start_column=dcf_interp_col,
                end_row=row_dcf_hdr,
                end_column=dcf_interp_col + 4,
            )
        except Exception:
            pass
        # Apply DCF interpretation formatting only on the DCF module rows,
        # not on the DCF sensitivity matrix (which also uses column K).
        for rr in range(row_dcf_hdr, row_dcf_eq + 1):
            dcf_interp = ws_local.cell(row=rr, column=dcf_interp_col)
            if dcf_interp.value in (None, ""):
                continue
            if rr <= row_dcf_eq:
                _force_merge_row(rr, dcf_interp_col, dcf_interp_col + 4)  # K:O for 216-222
            else:
                _merge_text_cell_horiz(rr, dcf_interp_col, dcf_interp_col + 4, max_span=4)
            dcf_interp.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            dcf_lines = _text_line_count(dcf_interp.value, _span_width_chars(rr, dcf_interp_col))
            row_target_heights[rr] = max(row_target_heights.get(rr, 0.0), _line_height(dcf_lines))

        # Keep Q-adjusted rows visually consistent.
        if row_qadj_ev in row_target_heights and row_qadj_yield in row_target_heights:
            row_target_heights[row_qadj_yield] = row_target_heights[row_qadj_ev]

        # Final DCF sensitivity alignment fix (ensure all gT columns match).
        for cc in range(dcf_label_col + 1, dcf_label_col + 6):  # H:L
            hcell = ws_local.cell(row=row_dcf_sens_hdr + 1, column=cc)
            hcell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
        for rr in range(row_dcf_sens_hdr + 2, row_dcf_sens_last_row + 1):
            ws_local.cell(row=rr, column=dcf_label_col).alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
            for cc in range(dcf_label_col + 1, dcf_label_col + 6):  # H:L
                c = ws_local.cell(row=rr, column=cc)
                c.alignment = Alignment(horizontal="right", vertical="bottom", wrap_text=False)

        for rr, hh in row_target_heights.items():
            ws_local.row_dimensions[rr].height = hh

        # Vertical separators: J|K only for the top panel; S|T through the market block.
        divider_side = Side(style="thick", color="000000")
        for rr in range(valuation_start_row, grid_start):
            target_cell = ws_local.cell(row=rr, column=10)  # J
            target_cell.border = Border(
                left=target_cell.border.left,
                right=divider_side,
                top=target_cell.border.top,
                bottom=target_cell.border.bottom,
            )
        for rr in range(valuation_start_row, row_market_hdr):
            target_cell = ws_local.cell(row=rr, column=19)  # S
            target_cell.border = Border(
                left=target_cell.border.left,
                right=divider_side,
                top=target_cell.border.top,
                bottom=target_cell.border.bottom,
            )

        # Additional requested separators.
        # Horizontal lines: above the lower stack and just above Trigger Scenarios.
        for rr in (grid_start - 1, row_scn_hdr - 1):
            end_cc = 19 if rr == grid_start - 1 else 15
            for cc in range(2, end_cc + 1):  # B..S or B..O
                c = ws_local.cell(row=rr, column=cc)
                c.border = Border(
                    left=c.border.left,
                    right=c.border.right,
                    top=c.border.top,
                    bottom=divider_side,
                )
        # Additional separators requested in market-implied area.
        for rr in (row_mi_toggle,):
            for cc in range(16, 22):  # P..U
                c = ws_local.cell(row=rr, column=cc)
                c.border = Border(
                    left=c.border.left,
                    right=c.border.right,
                    top=c.border.top,
                    bottom=divider_side,
                )
        for rr in range(row_market_hdr, min(row_convert_hdr, ws_local.max_row + 1)):
            c = ws_local.cell(row=rr, column=21)  # U
            c.border = Border(
                left=c.border.left,
                right=divider_side,
                top=c.border.top,
                bottom=c.border.bottom,
            )
        # Bottom line just above Convertible notes across B:U.
        for cc in range(2, 22):  # B..U
            c = ws_local.cell(row=row_convert_hdr - 1, column=cc)
            c.border = Border(
                left=c.border.left,
                right=c.border.right,
                top=c.border.top,
                bottom=divider_side,
            )
        # Vertical line between O and P for DCF/DCF-sensitivity rows.
        for rr in range(row_dcf_hdr, row_scn_hdr):
            c = ws_local.cell(row=rr, column=15)  # O
            c.border = Border(
                left=c.border.left,
                right=divider_side,
                top=c.border.top,
                bottom=c.border.bottom,
            )
        # Vertical line between K and L for market rows above Convertible notes.
        for rr in range(row_market_hdr, row_convert_hdr):
            c = ws_local.cell(row=rr, column=11)  # K
            c.border = Border(
                left=c.border.left,
                right=divider_side,
                top=c.border.top,
                bottom=c.border.bottom,
            )
        # Vertical line between K and L for lower valuation blocks.
        for rr in range(246, 262):
            c = ws_local.cell(row=rr, column=11)  # K
            c.border = Border(
                left=c.border.left,
                right=divider_side,
                top=c.border.top,
                bottom=c.border.bottom,
            )
        # Bottom line at row 261 across B:K.
        for cc in range(2, 12):  # B..K
            c = ws_local.cell(row=261, column=cc)
            c.border = Border(
                left=c.border.left,
                right=c.border.right,
                top=c.border.top,
                bottom=divider_side,
            )

        # 5) Tight spacing without moving block starts.
        for rr in sorted({row_scn_hdr - 1, row_toggle_hdr - 1, row_qadj_hdr - 1, row_dcf_hdr - 1, row_dcf_sens_hdr - 1}):
            if rr <= valuation_start_row:
                continue
            has_content = any(
                ws_local.cell(row=rr, column=cc).value not in (None, "")
                for cc in range(input_label_col, max(dcf_interp_col, market_interp_col) + 8)
            )
            if not has_content:
                ws_local.row_dimensions[rr].height = 13

        blank_run = 0
        for rr in range(valuation_start_row + 1, row_dcf_end + 1):
            has_content = any(
                ws_local.cell(row=rr, column=cc).value not in (None, "")
                for cc in range(input_label_col, max(dcf_interp_col, market_interp_col) + 8)
            )
            if has_content:
                blank_run = 0
                continue
            blank_run += 1
            if blank_run >= 2:
                ws_local.row_dimensions[rr].height = 13

    q_row_detect: Optional[int] = None
    for rr in range(1, min(300, ws.max_row) + 1):
        if str(ws.cell(row=rr, column=1).value or "").strip().lower() == "quarter":
            q_row_detect = rr
            break
    ws.freeze_panes = f"B{(q_row_detect + 1) if q_row_detect is not None else data_start_row}"
    if is_anf_profile:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and ("FY20" in cell.value or "Fiscal year" in cell.value):
                    cell.value = _anf_clean_visible_ui_text(cell.value)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = _updated_font(cell.font, size=font_size, bold=cell.font.b)
    for row_idx in [actuals_row, quarter_row]:
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = _updated_font(cell.font, size=header_size, bold=True)
    _record_writer_elapsed("write_excel.valuation.render.row_writes", row_write_elapsed)
    _record_writer_elapsed("write_excel.valuation.render.row_fills", row_fill_elapsed)
    _record_writer_substage("write_excel.valuation.render", valuation_render_started)
    # thick line separator (if forecast ever added, place after actuals)
    ws[f"{last_col_letter}{quarter_row}"].border = Border(right=thick)
    # Run valuation layout adjustments last (width/merge/spacing only).
    _apply_valuation_layout(ws)
    if is_anf_profile:
        _anf_clear_valuation_side_panels(ws, start_col=15, end_col=29)
        _write_anf_valuation_side_panel(ws)

    valuation_start_row_result = valuation_header_row
    for rr in range(1, min(ws.max_row, 400) + 1):
        if str(ws.cell(row=rr, column=2).value or "").strip().lower() == "valuation":
            valuation_start_row_result = rr
            break
    return ValuationFinalLayoutResult(
        freeze_panes=str(ws.freeze_panes or ""),
        valuation_start_row=valuation_start_row_result,
        max_row=int(ws.max_row),
        max_column=int(ws.max_column),
    )
