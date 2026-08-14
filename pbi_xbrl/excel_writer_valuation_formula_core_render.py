"""Formula core render adapter for the Valuation worksheet."""
from __future__ import annotations

import copy
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, List, MutableMapping, Optional

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


@dataclass(frozen=True)
class ValuationFormulaCoreRenderDeps:
    runtime: MutableMapping[str, Any]


@dataclass(frozen=True)
class ValuationFormulaCoreRenderResult:
    valuation_header_row: int
    valuation_inputs_row: int
    input_label_col: int
    input_value_col: int
    input_basis_col: int
    input_hint_col: int
    output_label_col: int
    output_value_col: int
    output_interp_col: int
    market_label_col: int
    market_value_col: int
    market_interp_col: int
    scn_label_col: int
    scn_value_col: int
    scn_interp_col: int
    driver_label_col: int
    driver_value_col: int
    toggle_label_col: int
    toggle_value_col: int
    qadj_label_col: int
    qadj_value_col: int
    qadj_text_col: int
    dcf_label_col: int
    dcf_value_col: int
    dcf_interp_col: int
    grid_start: int
    grid_layout_width: int
    right_stack_anchor: int
    date_ref: Any
    row_price: int
    row_asof: int
    row_shares_out: int
    row_shares_dil: int
    row_net_debt: int
    row_ebitda_ttm: int
    row_adj_ebitda_ttm: int
    row_fcf_ttm: int
    row_adj_fcf_ttm: int
    row_rev_ttm: int
    row_eps_ttm: int
    row_adj_eps_ttm: int
    row_bv: int
    row_tbv: int
    row_tgt_ev_adj: int
    row_tgt_ev: int
    row_tgt_fcf: int
    row_capex_ttm: int
    row_int_paid_ttm: int
    row_owner_maint_ratio: int
    row_owner_recurring: int
    row_owner_wc_norm: int
    row_share_mode: int
    row_out_hdr: int
    row_mktcap: int
    row_ev: int
    row_implied_ev_adj: int
    row_implied_ev: int
    row_fcff_proxy_ttm: int
    row_implied_fcff: int
    row_equity_fcf: int
    row_owner_fcf_ttm: int
    row_owner_fcf_yield: int
    row_eq_adj: int
    row_eq_ev: int
    row_eq_fcf: int
    row_pe: int
    row_pe_adj: int
    row_ev_sales: int
    row_pb: int
    row_ptbv: int
    row_mi_hdr: int
    row_mi_market_ev: int
    row_mi_dcf_ev: int
    row_mi_curr_wacc: int
    row_mi_curr_gt: int
    row_mi_tbl_hdr: int
    row_mi_wacc_start: int
    row_mi_wacc_end: int
    row_mi_toggle: int
    row_dcf_hdr: int
    row_dcf_start: int
    row_dcf_g: int
    row_dcf_gt: int
    row_dcf_wacc: int
    row_dcf_ev: int
    row_dcf_eq: int
    row_dcf_sens_hdr: int
    row_dcf_sens_last_row: int
    row_scn_hdr: int
    row_scn_profile: int
    row_scn_growth: int
    row_scn_margin: int
    row_scn_refi: int
    row_scn_buyback: int
    row_scn_adj_ebitda: int
    row_scn_owner_fcf: int
    row_scn_eq_ev: int
    row_scn_eq_fcf: int
    row_market_hdr: int
    row_req_adj_ebitda: int
    row_req_adj_delta: int
    row_req_fcff: int
    row_req_fcff_delta: int
    row_req_owner_fcf: int
    row_req_owner_delta: int
    row_qa: int
    row_drv_hdr: int
    row_drv_rev: int
    row_drv_margin: int
    row_drv_fcf: int
    row_drv_lev: int
    row_toggle_hdr: int
    row_toggle_reg: int
    row_toggle_gap: int
    row_toggle_lev: int
    row_toggle_conc: int
    row_qadj_hdr: int
    row_qadj_ev_adj: int
    row_qadj_ev: int
    row_qadj_yield: int
    row_dcf_end: int
    row_hv_hdr: int
    row_hv_total: int
    row_hv_prof: int
    row_hv_cash: int
    row_hv_delev: int
    row_hv_quality: int
    row_hv_narr: int
    row_hv_b1: int
    row_hv_b2: int
    row_hv_b3: int
    row_hv_b4: int
    row_hv_b5: int
    row_convert_hdr: int
    convert_header_end_col: int
    qa_msgs: List[str]
    tieout_diff_m: Optional[float]
    fair_denom: str
    normalize_thesis_bridge_basis: Callable[..., Any]
    set_formula_name: Callable[..., Any]
    set_cell_comment: Callable[..., Any]


def render_valuation_formula_core(
    deps: ValuationFormulaCoreRenderDeps,
) -> ValuationFormulaCoreRenderResult:
    __rt = deps.runtime
    context_globals = dict(__rt.get("context_globals") or {})

    def _rt_get(name: str) -> Any:
        if name in __rt:
            return __rt[name]
        if name in context_globals:
            return context_globals[name]
        return globals().get(name)

    Alignment = _rt_get("Alignment")
    Border = _rt_get("Border")
    CellIsRule = _rt_get("CellIsRule")
    DefinedName = _rt_get("DefinedName")
    Font = _rt_get("Font")
    Path = _rt_get("Path")
    PatternFill = _rt_get("PatternFill")
    Side = _rt_get("Side")
    _collapse_repeated_leading_ngram_local = _rt_get("_collapse_repeated_leading_ngram_local")
    _dedupe_canonical_text_parts_local = _rt_get("_dedupe_canonical_text_parts_local")
    _htmlish_to_text = _rt_get("_htmlish_to_text")
    _quarter_notes_view = _rt_get("_quarter_notes_view")
    _resolve_col = _rt_get("_resolve_col")
    _safe_text_value = _rt_get("_safe_text_value")
    _set_cell_comment_local = _rt_get("_set_cell_comment_local")
    _source_backed_debt_tranches_from_slides = _rt_get("_source_backed_debt_tranches_from_slides")
    adj_ebitda_map = _rt_get("adj_ebitda_map")
    adj_ebitda_ttm_map = _rt_get("adj_ebitda_ttm_map")
    adj_eps_ttm_map = _rt_get("adj_eps_ttm_map")
    adj_fcf_ttm_map = _rt_get("adj_fcf_ttm_map")
    bold = _rt_get("bold")
    bv_share_map = _rt_get("bv_share_map")
    capex_map = _rt_get("capex_map")
    capex_ttm_map = _rt_get("capex_ttm_map")
    cash_map = _rt_get("cash_map")
    cfo_map = _rt_get("cfo_map")
    company_overview = _rt_get("company_overview")
    copy = _rt_get("copy")
    debt_core_map = _rt_get("debt_core_map")
    debt_tranches_latest = _rt_get("debt_tranches_latest")
    ebitda_map = _rt_get("ebitda_map")
    ebitda_ttm_map = _rt_get("ebitda_ttm_map")
    font_size = _rt_get("font_size")
    get_column_letter = _rt_get("get_column_letter")
    glx_normalize_text = _rt_get("glx_normalize_text")
    header_fill = _rt_get("header_fill")
    input_fill = _rt_get("input_fill")
    int_paid_ttm_map = _rt_get("int_paid_ttm_map")
    is_anf_profile = _rt_get("is_anf_profile")
    last4_quarters_map = _rt_get("last4_quarters_map")
    last_col_letter = _rt_get("last_col_letter")
    net_income_map = _rt_get("net_income_map")
    net_lev_map = _rt_get("net_lev_map")
    owner_maint_capex_ratio_default = _rt_get("owner_maint_capex_ratio_default")
    pd = _rt_get("pd")
    price = _rt_get("price")
    qn_compact_snippet = _rt_get("qn_compact_snippet")
    qs = _rt_get("qs")
    quarter_columns = _rt_get("quarter_columns")
    quarter_notes = _rt_get("quarter_notes")
    re = _rt_get("re")
    rev_map = _rt_get("rev_map")
    rev_ttm_map = _rt_get("rev_ttm_map")
    row_operating_margin_pct = _rt_get("row_operating_margin_pct")
    row_operating_margin_ttm_pct = _rt_get("row_operating_margin_ttm_pct")
    section_fill = _rt_get("section_fill")
    shares_for_value_map = _rt_get("shares_for_value_map")
    shares_map = _rt_get("shares_map")
    shares_out_map = _rt_get("shares_out_map")
    slides_debt = _rt_get("slides_debt")
    tbv_share_map = _rt_get("tbv_share_map")
    thin_border = _rt_get("thin_border")
    ticker = _rt_get("ticker")
    tieout_diff_m = _rt_get("tieout_diff_m")
    valuation_grid_df = _rt_get("valuation_grid_df")
    valuation_price_input_available = _rt_get("valuation_price_input_available")
    wb = _rt_get("wb")
    ws = _rt_get("ws")

    # Valuation boxes
    # Keep numeric value cells narrow; make text boxes wider by spanning label columns.
    input_label_col = 2    # B
    input_value_col = 4    # D
    input_basis_col = 5    # E
    input_hint_col = 6     # F
    # Valuation panel layout (rows 193+ only).
    output_label_col = 11  # K
    output_value_col = 14  # N
    output_interp_col = 15 # O
    market_label_col = 12  # L
    market_value_col = 16  # P
    market_interp_col = 17 # Q
    scn_label_col = 2      # B
    scn_value_col = 5      # E
    scn_interp_col = 6     # F
    driver_label_col = 2   # B
    driver_value_col = 4   # D
    toggle_label_col = 2   # B
    toggle_value_col = 5   # E
    qadj_label_col = 2     # B
    qadj_value_col = 5     # E
    qadj_text_col = 6      # F
    grid_layout_width = int(
        max(
            4,
            min(
                4,
                len(list(valuation_grid_df.columns))
            )
            if valuation_grid_df is not None and not valuation_grid_df.empty
            else 4,
        )
    )
    # DCF block starts at column G by design (layout-only; no valuation logic impact).
    dcf_label_col = 7
    dcf_value_col = dcf_label_col + 3
    dcf_interp_col = dcf_value_col + 1  # K
    if dcf_interp_col >= market_label_col:
        shift_cols = dcf_interp_col - market_label_col + 1
        market_label_col += shift_cols
        market_value_col += shift_cols
        market_interp_col += shift_cols
    input_label_col_letter = get_column_letter(input_label_col)
    input_basis_col_letter = get_column_letter(input_basis_col)
    input_hint_col_letter = get_column_letter(input_hint_col)
    output_label_col_letter = get_column_letter(output_label_col)
    output_interp_col_letter = get_column_letter(output_interp_col)
    valuation_header_row = 192
    valuation_inputs_row = valuation_header_row + 1
    valuation_shift = valuation_header_row - 2
    ws[f"{input_label_col_letter}{valuation_header_row}"] = "Valuation"
    ws[f"{input_label_col_letter}{valuation_header_row}"].font = bold
    ws[f"{input_label_col_letter}{valuation_header_row}"].fill = section_fill
    try:
        ws.merge_cells(
            start_row=valuation_header_row,
            start_column=input_label_col,
            end_row=valuation_header_row,
            end_column=19,  # S
        )
    except Exception:
        pass
    ws[f"{input_label_col_letter}{valuation_inputs_row}"] = "Inputs"
    ws[f"{input_label_col_letter}{valuation_inputs_row}"].font = bold
    ws[f"{input_label_col_letter}{valuation_inputs_row}"].fill = header_fill
    ws[f"{input_basis_col_letter}{valuation_inputs_row}"] = "Basis"
    ws[f"{input_basis_col_letter}{valuation_inputs_row}"].font = bold
    ws[f"{input_basis_col_letter}{valuation_inputs_row}"].fill = header_fill
    ws[f"{input_hint_col_letter}{valuation_inputs_row}"] = "Hint"
    ws[f"{input_hint_col_letter}{valuation_inputs_row}"].font = bold
    ws[f"{input_hint_col_letter}{valuation_inputs_row}"].fill = header_fill
    ws[f"{output_label_col_letter}{valuation_inputs_row}"] = "Outputs"
    ws[f"{output_label_col_letter}{valuation_inputs_row}"].font = bold
    ws[f"{output_label_col_letter}{valuation_inputs_row}"].fill = header_fill
    ws[f"{output_interp_col_letter}{valuation_inputs_row}"] = "Interpretation"
    ws[f"{output_interp_col_letter}{valuation_inputs_row}"].font = bold
    ws[f"{output_interp_col_letter}{valuation_inputs_row}"].fill = header_fill

    def _box(
        row_idx: int,
        label: str,
        value: Optional[str] = None,
        is_input: bool = False,
        number_format: Optional[str] = None,
        label_col: Optional[int] = None,
        value_col: Optional[int] = None,
    ) -> None:
        lc = input_label_col if label_col is None else int(label_col)
        vc = input_value_col if value_col is None else int(value_col)
        if vc - lc > 1:
            try:
                ws.merge_cells(start_row=row_idx, start_column=lc, end_row=row_idx, end_column=vc - 1)
            except Exception:
                pass
        ws.cell(row=row_idx, column=lc, value=label).font = Font(size=font_size, bold=False)
        vcell = ws.cell(row=row_idx, column=vc, value=value)
        if is_input:
            vcell.fill = input_fill
        if number_format:
            vcell.number_format = number_format

    latest_col = last_col_letter
    date_ref = pd.Timestamp(qs[-1]) if qs else None
    # input/output row indices
    row_price = 4 + valuation_shift
    row_asof = 5 + valuation_shift
    row_shares_out = 6 + valuation_shift
    row_shares_dil = 7 + valuation_shift
    row_net_debt = 8 + valuation_shift
    row_ebitda_ttm = 9 + valuation_shift
    row_adj_ebitda_ttm = 10 + valuation_shift
    row_fcf_ttm = 11 + valuation_shift
    row_adj_fcf_ttm = 12 + valuation_shift
    row_rev_ttm = 13 + valuation_shift
    row_eps_ttm = 14 + valuation_shift
    row_adj_eps_ttm = 15 + valuation_shift
    row_bv = 16 + valuation_shift
    row_tbv = 17 + valuation_shift
    row_tgt_ev_adj = 18 + valuation_shift
    row_tgt_ev = 19 + valuation_shift
    row_tgt_fcf = 20 + valuation_shift
    row_capex_ttm = 21 + valuation_shift
    row_int_paid_ttm = 22 + valuation_shift
    row_owner_maint_ratio = 23 + valuation_shift
    row_owner_recurring = 24 + valuation_shift
    row_owner_wc_norm = 25 + valuation_shift
    row_share_mode = 26 + valuation_shift
    row_out_hdr = valuation_header_row + 1
    row_mktcap = row_out_hdr + 1
    row_ev = row_out_hdr + 2
    row_implied_ev_adj = row_out_hdr + 3
    row_implied_ev = row_out_hdr + 4
    row_fcff_proxy_ttm = row_out_hdr + 5
    row_implied_fcff = row_out_hdr + 6
    row_equity_fcf = row_out_hdr + 7
    row_owner_fcf_ttm = row_out_hdr + 8
    row_owner_fcf_yield = row_out_hdr + 9
    row_eq_adj = row_out_hdr + 10
    row_eq_ev = row_out_hdr + 11
    row_eq_fcf = row_out_hdr + 12
    row_pe = row_out_hdr + 13
    row_pe_adj = row_out_hdr + 14
    row_ev_sales = row_out_hdr + 15
    row_pb = row_out_hdr + 16
    row_ptbv = row_out_hdr + 17
    dcf_sens_g_vals = [0.00, 0.01, 0.02, 0.03, 0.04]
    dcf_sens_wacc_vals = [0.08, 0.09, 0.10, 0.11, 0.12, 0.13, 0.14, 0.15]
    # Lower valuation stack anchors below the top valuation panel.
    grid_start = max(row_share_mode, row_ptbv) + 2
    right_stack_anchor = grid_start - 1
    # Market-implied terminal g block.
    row_mi_hdr = right_stack_anchor
    row_mi_market_ev = row_mi_hdr + 1
    row_mi_dcf_ev = row_mi_hdr + 2
    row_mi_curr_wacc = row_mi_hdr + 3
    row_mi_curr_gt = row_mi_hdr + 4
    row_mi_tbl_hdr = row_mi_hdr + 5
    row_mi_wacc_start = row_mi_hdr + 6
    row_mi_wacc_end = row_mi_wacc_start + len(dcf_sens_wacc_vals) - 1
    row_mi_toggle = row_mi_wacc_end + 4
    # DCF block rendered to the right of valuation sensitivity grid.
    row_dcf_hdr = right_stack_anchor
    row_dcf_start = row_dcf_hdr + 1
    row_dcf_g = row_dcf_hdr + 2
    row_dcf_gt = row_dcf_hdr + 3
    row_dcf_wacc = row_dcf_hdr + 4
    row_dcf_ev = row_dcf_hdr + 5
    row_dcf_eq = row_dcf_hdr + 6
    # Keep one compact spacer row between DCF module and DCF sensitivity matrix.
    row_dcf_sens_hdr = row_dcf_eq + 2
    row_dcf_sens_last_row = row_dcf_sens_hdr + 1 + len(dcf_sens_wacc_vals)
    # Keep Scenario block one row below the last WACC row.
    row_scn_hdr = row_dcf_sens_last_row + 1
    row_scn_profile = row_scn_hdr + 1
    row_scn_growth = row_scn_hdr + 2
    row_scn_margin = row_scn_hdr + 3
    row_scn_refi = row_scn_hdr + 4
    row_scn_buyback = row_scn_hdr + 5
    row_scn_adj_ebitda = row_scn_hdr + 6
    row_scn_owner_fcf = row_scn_hdr + 7
    row_scn_eq_ev = row_scn_hdr + 8
    row_scn_eq_fcf = row_scn_hdr + 9
    # What Market Is Pricing block aligns with the trigger section.
    row_market_hdr = row_scn_hdr
    row_req_adj_ebitda = row_market_hdr + 1
    row_req_adj_delta = row_market_hdr + 2
    row_req_fcff = row_market_hdr + 3
    row_req_fcff_delta = row_market_hdr + 4
    row_req_owner_fcf = row_market_hdr + 5
    row_req_owner_delta = row_market_hdr + 6
    row_qa = row_scn_eq_fcf + 1
    # Move supporting blocks down together to avoid overlap with expanded DCF sensitivity.
    row_drv_hdr = row_scn_eq_fcf + 2
    row_drv_rev = row_drv_hdr + 1
    row_drv_margin = row_drv_hdr + 2
    row_drv_fcf = row_drv_hdr + 3
    row_drv_lev = row_drv_hdr + 4
    row_toggle_hdr = row_drv_lev + 2
    row_toggle_reg = row_toggle_hdr + 1
    row_toggle_gap = row_toggle_hdr + 2
    row_toggle_lev = row_toggle_hdr + 3
    row_toggle_conc = row_toggle_hdr + 4
    row_qadj_hdr = row_toggle_conc + 2
    row_qadj_ev_adj = row_qadj_hdr + 1
    row_qadj_ev = row_qadj_hdr + 2
    row_qadj_yield = row_qadj_hdr + 3
    row_dcf_end = max(row_dcf_sens_last_row, row_qadj_yield, row_mi_toggle)
    row_hv_hdr = 68
    row_hv_total = 69
    row_hv_prof = 70
    row_hv_cash = 71
    row_hv_delev = 72
    row_hv_quality = 73
    row_hv_narr = 74
    row_hv_b1 = 76
    row_hv_b2 = 77
    row_hv_b3 = 78
    row_hv_b4 = 79
    row_hv_b5 = 80

    def _set_named_range(name: str, row_idx: int, value_col: Optional[int] = None) -> None:
        vc = input_value_col if value_col is None else int(value_col)
        ref = f"'Valuation'!${get_column_letter(vc)}${row_idx}"
        try:
            if name in wb.defined_names:
                del wb.defined_names[name]
        except Exception:
            pass
        try:
            wb.defined_names.add(DefinedName(name=name, attr_text=ref))
        except Exception:
            # fallback for older openpyxl API
            try:
                wb.defined_names.append(DefinedName(name=name, attr_text=ref))
            except Exception:
                pass

    def _set_formula_name(name: str, formula: str) -> None:
        ftxt = str(formula or "").strip()
        if not ftxt:
            return
        if ftxt.startswith("="):
            ftxt = ftxt[1:].strip()
        try:
            if name in wb.defined_names:
                del wb.defined_names[name]
        except Exception:
            pass
        try:
            wb.defined_names.add(DefinedName(name=name, attr_text=ftxt))
        except Exception:
            try:
                wb.defined_names.append(DefinedName(name=name, attr_text=ftxt))
            except Exception:
                pass

    def _set_named_range_ref(name: str, ref: str) -> None:
        if not ref:
            return
        try:
            if name in wb.defined_names:
                del wb.defined_names[name]
        except Exception:
            pass
        try:
            wb.defined_names.add(DefinedName(name=name, attr_text=ref))
        except Exception:
            try:
                wb.defined_names.append(DefinedName(name=name, attr_text=ref))
            except Exception:
                pass

    def _wrap_iferror(formula: str) -> str:
        ftxt = str(formula or "").strip()
        if not ftxt:
            return ftxt
        if not ftxt.startswith("="):
            return ftxt
        inner = ftxt[1:].strip()
        inner_up = inner.upper()
        if inner_up.startswith("IFERROR(") or inner_up.startswith("OMFEL("):
            return ftxt
        return f'=IFERROR({inner},"")'

    def _set_cell_comment(cell: Any, text: str) -> None:
        if not text:
            return
        try:
            _set_cell_comment_local(cell, text)
        except Exception:
            pass

    def _first_sentence_or_trunc(text: str, max_chars: int = 110) -> str:
        t = re.sub(r"\s+", " ", str(text or "")).strip()
        if not t:
            return ""
        sent = re.search(r"[.!?](?:\s|$)", t)
        if sent and sent.end() <= max_chars:
            return t[: sent.end()].strip()
        if len(t) <= max_chars:
            return t
        cut = t[:max_chars]
        ws_cut = cut.rfind(" ")
        if ws_cut >= int(max_chars * 0.6):
            cut = cut[:ws_cut]
        return cut.rstrip(" ,;:-") + "..."

    def _collapse_leading_repeated_phrase_local(text_in: Any) -> str:
        txt = re.sub(r"\s+", " ", str(text_in or "")).strip()
        if not txt:
            return ""
        words = txt.split()
        max_window = min(12, len(words) // 2)
        for size in range(max_window, 1, -1):
            lhs = " ".join(words[:size]).strip()
            rhs = " ".join(words[size : size * 2]).strip()
            if lhs and rhs and lhs.lower() == rhs.lower():
                rest = " ".join(words[size * 2 :]).strip()
                return f"{lhs} {rest}".strip() if rest else lhs
        return txt

    def _extract_driver_delta_basis_local(text_in: str, *, default_basis: str = "YoY") -> str:
        txt = glx_normalize_text(str(text_in or ""))
        if re.search(r"\bQoQ delta\b|\bquarter[- ]over[- ]quarter\b|\bvs\.?\s+prior quarter\b", txt, re.I):
            return "QoQ"
        if re.search(r"\bYoY delta\b|\byear[- ]over[- ]year\b|\bvs\.?\s+prior year\b", txt, re.I):
            return "YoY"
        return default_basis

    def _clean_visible_driver_text_local(text_in: Any, max_chars: int = 110) -> str:
        txt = glx_normalize_text(str(text_in or ""))
        if not txt:
            return ""
        parts: List[str] = []
        seen_parts: set[str] = set()
        for raw_part in re.split(r"\s*\|\s*", txt):
            part = _collapse_leading_repeated_phrase_local(glx_normalize_text(raw_part))
            if not part:
                continue
            part_key = part.lower()
            if part_key in seen_parts:
                continue
            seen_parts.add(part_key)
            parts.append(part)
        txt = " | ".join(parts) if parts else txt
        txt = _collapse_leading_repeated_phrase_local(txt)
        if re.search(r"\brevenue ttm still under pressure\b", txt, re.I):
            yoy_match = re.search(r"Revenue TTM YoY\s+-?[0-9]+(?:\.[0-9]+)?%", txt, re.I)
            return (
                f"Revenue TTM still under pressure; {yoy_match.group(0)}"
                if yoy_match
                else "Revenue TTM still under pressure"
            )
        if re.search(r"\bFCF TTM accelerated\b", txt, re.I):
            basis = _extract_driver_delta_basis_local(txt, default_basis="YoY")
            delta_match = re.search(r"(?:YoY|QoQ)?\s*delta\s+\$-?[0-9]+(?:\.[0-9]+)?m", txt, re.I)
            if delta_match:
                delta_txt = re.sub(r"^(?:YoY|QoQ)?\s*", "", delta_match.group(0), flags=re.I).strip()
                return f"FCF TTM accelerated; {basis} {delta_txt}"
            return "FCF TTM accelerated"
        if re.search(r"\bNet debt (declined|increased)\b", txt, re.I):
            basis = _extract_driver_delta_basis_local(txt, default_basis="YoY")
            direction_match = re.search(r"\bNet debt (declined|increased)\b", txt, re.I)
            direction_txt = (
                f"Net debt {str(direction_match.group(1) or '').lower()}"
                if direction_match
                else "Net debt"
            )
            delta_match = re.search(r"Net debt delta\s+\$-?[0-9]+(?:\.[0-9]+)?m", txt, re.I)
            if delta_match:
                delta_amt = re.sub(r"^Net debt delta\s+", "", delta_match.group(0), flags=re.I).strip()
                return f"{direction_txt}; {basis} delta {delta_amt}"
            return direction_txt
        if re.search(r"\bEBITDA margin (expanded|compressed)\b", txt, re.I):
            basis = _extract_driver_delta_basis_local(txt, default_basis="YoY")
            direction_match = re.search(r"\bEBITDA margin (expanded|compressed)\b", txt, re.I)
            direction_txt = (
                f"EBITDA margin {str(direction_match.group(1) or '').lower()}"
                if direction_match
                else "EBITDA margin"
            )
            delta_match = re.search(r"(?:delta\s+)?([+-]?\d+(?:\.\d+)?)\s*bps", txt, re.I)
            if delta_match:
                delta_num = str(delta_match.group(1) or "").strip()
                if delta_num and not delta_num.startswith(("+", "-")):
                    delta_num = f"+{delta_num}" if "expanded" in direction_txt.lower() else f"-{delta_num}"
                return f"{direction_txt}; {basis} delta {delta_num} bps"
            return direction_txt
        if re.search(r"\bRevolver utilization notable\b", txt, re.I):
            return "Revolver utilization notable"
        txt = re.sub(r"\b(N/A)(?:\s+\1)+\b", r"\1", txt, flags=re.I)
        return _first_sentence_or_trunc(txt, max_chars=max_chars)

    def _clean_audit_excerpt_local(text_in: Any, max_chars: int = 320) -> str:
        txt = glx_normalize_text(str(text_in or ""))
        if not txt:
            return ""
        parts: List[str] = []
        seen_parts: set[str] = set()
        for raw_part in re.split(r"\s*\|\s*", txt):
            part = _collapse_leading_repeated_phrase_local(glx_normalize_text(raw_part))
            if re.search(r"\b(FCF TTM accelerated|Revenue TTM still under pressure|Net debt (?:declined|increased)|EBITDA margin (?:expanded|compressed)|Revolver utilization notable)\b", part, re.I):
                part = _clean_visible_driver_text_local(part, max_chars=min(max_chars, 160))
            if not part:
                continue
            part_key = part.lower()
            if part_key in seen_parts or any((len(part_key) >= 24 and (part_key in seen or seen in part_key)) for seen in seen_parts):
                continue
            seen_parts.add(part_key)
            parts.append(part)
        txt = " | ".join(parts) if parts else txt
        txt = _collapse_leading_repeated_phrase_local(txt)
        if len(txt) <= max_chars:
            return txt
        kept_parts: List[str] = []
        cur_len = 0
        for part in parts:
            projected = cur_len + (3 if kept_parts else 0) + len(part)
            if projected > max_chars:
                break
            kept_parts.append(part)
            cur_len = projected
        if kept_parts:
            return " | ".join(kept_parts).rstrip(" ,;:-") + ("..." if len(" | ".join(kept_parts)) < len(txt) else "")
        return _first_sentence_or_trunc(txt, max_chars=max_chars)

    def _canonicalize_audit_excerpt_local(text_in: Any) -> str:
        txt = _clean_audit_excerpt_local(text_in, max_chars=1200)
        if not txt:
            return ""
        txt = re.sub(r"\s+", " ", txt).strip(" |")
        txt = re.sub(r"\s*([|,:;])\s*", r" \1 ", txt)
        txt = re.sub(r"\s+", " ", txt).strip()
        parts = _dedupe_canonical_text_parts_local(re.split(r"\s*\|\s*", txt))
        txt = " | ".join(parts) if parts else txt
        prior_txt = None
        while txt and txt != prior_txt:
            prior_txt = txt
            txt = re.sub(
                r"\b([A-Za-z][A-Za-z0-9/%$().,\- ]{3,90}?)\s+\1\b",
                r"\1",
                txt,
                flags=re.I,
            )
            txt = re.sub(
                r"(?i)\b([^|]{4,120}?)\s*\|\s*\1\b",
                r"\1",
                txt,
            )
            txt = re.sub(
                r"^\s*([^|]{4,80}?)\s+\1\b",
                r"\1",
                txt,
                flags=re.I,
            )
            txt = _collapse_repeated_leading_ngram_local(txt)
        parts = _dedupe_canonical_text_parts_local(re.split(r"\s*\|\s*", txt))
        txt = " | ".join(parts) if parts else txt
        return glx_normalize_text(txt)

    def _audit_doc_family_local(source_doc_in: Any) -> str:
        source_doc = str(source_doc_in or "").strip()
        if not source_doc:
            return ""
        try:
            name = Path(source_doc).name
        except Exception:
            name = source_doc.replace("\\", "/").split("/")[-1]
        name = glx_normalize_text(name).lower()
        name = re.sub(r"\.[a-z0-9]+$", "", name)
        name = re.sub(r"^doc_\d+_", "", name)
        name = re.sub(r"[-_]+", "_", name)
        return name

    def _quarter_notes_audit_canonical_rows_local(rows_in: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if not rows_in:
            return []

        def _canonical_source_group_local(row_in: Dict[str, Any]) -> str:
            source_type = str(row_in.get("source_type") or "").strip().lower()
            idea_label = glx_normalize_text(
                str(
                    row_in.get("idea_label")
                    or row_in.get("metric_display")
                    or row_in.get("family")
                    or row_in.get("candidate_type")
                    or ""
                )
            ).lower()
            if source_type == "model_metric":
                return f"model_metric:{idea_label or 'unknown'}"
            doc_family = _audit_doc_family_local(row_in.get("source_doc"))
            if doc_family:
                return f"{source_type or 'source'}:{doc_family}"
            return f"{source_type or 'source'}:unknown"

        def _stage_priority_local(stage_in: Any) -> int:
            stage_low = str(stage_in or "").strip().lower()
            stage_rank = {
                "saved_workbook_visible": 90,
                "selection_kept": 80,
                "quality_filtered": 70,
                "routed_to_bucket": 60,
                "score_assigned": 50,
                "render_summary_generated": 40,
                "candidate_created": 30,
                "source_detected": 20,
                "quality_review": 10,
            }
            return stage_rank.get(stage_low, 0)

        grouped: Dict[Tuple[str, str, str, str], List[Tuple[int, Dict[str, Any]]]] = {}
        for idx, raw_row in enumerate(rows_in):
            row = dict(raw_row)
            row["source_excerpt"] = _canonicalize_audit_excerpt_local(row.get("source_excerpt"))
            row["normalized_source_doc_family"] = _audit_doc_family_local(row.get("source_doc"))
            row["canonical_source_group"] = _canonical_source_group_local(row)
            idea_label = glx_normalize_text(
                str(row.get("idea_label") or row.get("metric_display") or row.get("family") or "")
            ).lower()
            key = (
                str(row.get("quarter") or ""),
                idea_label,
                str(row.get("source_excerpt") or "").lower(),
                str(row.get("canonical_source_group") or ""),
            )
            grouped.setdefault(key, []).append((idx, row))

        rows_out: List[Tuple[int, Dict[str, Any]]] = []
        for _, group_rows in grouped.items():
            best_idx, best_row = max(
                group_rows,
                key=lambda pair: (
                    _stage_priority_local(pair[1].get("stage")),
                    int(float(pd.to_numeric(pair[1].get("score_total"), errors="coerce") or 0.0)),
                    len(str(pair[1].get("final_summary") or "")),
                    -pair[0],
                ),
            )
            group_only_rows = [dict(row) for _, row in group_rows]
            merged = dict(best_row)
            merged["support_count"] = len(group_only_rows)
            merged["source_count"] = len(
                {
                    str(row.get("source_doc") or row.get("normalized_source_doc_family") or row.get("canonical_source_group") or "")
                    for row in group_only_rows
                }
            )
            if not str(merged.get("final_summary") or "").strip():
                for row in group_only_rows:
                    final_summary = str(row.get("final_summary") or "").strip()
                    if final_summary:
                        merged["final_summary"] = final_summary
                        break
            if str(merged.get("final_summary") or "").strip():
                merged["final_summary"] = _canonicalize_audit_excerpt_local(merged.get("final_summary"))
            rows_out.append((min(idx for idx, _ in group_rows), merged))

        rows_out.sort(key=lambda pair: pair[0])
        return [row for _, row in rows_out]

    def _set_snippet_with_comment(
        cell: Any,
        full_text: str,
        max_chars: int = 110,
        extra_comment: str = "",
        visible_text: str = "",
    ) -> None:
        txt = str(full_text or "").strip()
        visible = str(visible_text or txt).strip()
        cell.value = re.sub(r"\s+", " ", visible) if visible else ""
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if txt:
            cmt = txt if not extra_comment else f"{txt}\n\n{extra_comment}"
            _set_cell_comment(cell, cmt)

    def _set_input_meta(row_idx: int, basis: str, hint: str) -> None:
        bcell = ws.cell(row=row_idx, column=input_basis_col, value=basis)
        bcell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        hcell = ws.cell(row=row_idx, column=input_hint_col)
        _set_snippet_with_comment(hcell, hint, max_chars=320)
        hcell.font = Font(size=max(8, font_size - 1))

    def _set_interpretation(row_idx: int, text: str, interp_col: int) -> None:
        cell = ws.cell(row=row_idx, column=interp_col)
        _set_snippet_with_comment(cell, text, max_chars=420)
        cell.font = Font(size=max(8, font_size - 1))

    def _normalize_thesis_bridge_basis(name_in: str, value_in: Any) -> Any:
        num = pd.to_numeric(value_in, errors="coerce")
        if pd.isna(num):
            return value_in
        val = float(num)
        key = str(name_in or "").strip().lower()
        if key in {"target_ev_adjebitda", "target_ev_yield", "maintcapexratio"}:
            return val
        if key in {"shares", "sharesdiluted"}:
            return val / 1e6 if abs(val) >= 100000.0 else val
        if key in {
            "thesisbaseadjebitda_fy",
            "netdebt",
            "base_ebitda",
            "adj_ebitda",
            "fcf_ttm",
            "adj_fcf_ttm",
            "revenue_ttm",
            "capex_ttm",
            "interestpaid_ttm",
            "recurringcashcosts",
            "wcnormalization",
        }:
            return val / 1e6 if abs(val) >= 100000.0 else val
        return val

    # Mini usage box intentionally omitted here to avoid overlapping existing top narrative rows.

    _box(row_price, "Price", price, True, "$#,##0.00")
    price_cell = ws.cell(row=row_price, column=input_value_col)
    price_cell.font = Font(color="0070C0", bold=True, size=font_size)
    _box(row_asof, "As of", pd.Timestamp(qs[-1]).date() if qs else None, False, "yyyy-mm-dd")
    if date_ref is not None:
        sh_out = shares_out_map.get(date_ref)
        if sh_out is None:
            sh_out = shares_for_value_map.get(date_ref)
        sh_dil = shares_map.get(date_ref)
        if sh_dil is None:
            sh_dil = sh_out
        nd = None
        if debt_core_map.get(date_ref) is not None and cash_map.get(date_ref) is not None:
            nd = debt_core_map.get(date_ref) - cash_map.get(date_ref)
        _box(row_shares_out, "Shares outstanding (m)", (sh_out / 1e6) if sh_out is not None else None, False, "#,##0.000")
        _box(row_shares_dil, "Shares diluted (m)", (sh_dil / 1e6) if sh_dil is not None else None, False, "#,##0.000")
        _box(row_net_debt, "Net debt (core, $m)", (nd / 1e6) if nd is not None else None, False, "#,##0.000")
        _box(row_ebitda_ttm, "EBITDA TTM ($m)", (ebitda_ttm_map.get(date_ref) / 1e6) if ebitda_ttm_map.get(date_ref) is not None else None, False, "#,##0.000")
        adj_ttm_val = adj_ebitda_ttm_map.get(date_ref) if adj_ebitda_ttm_map else None
        _box(row_adj_ebitda_ttm, "Adj EBITDA TTM ($m)", (adj_ttm_val / 1e6) if adj_ttm_val is not None else None, False, "#,##0.000")
        fcf_ttm = None
        # compute FCF TTM from last 4 quarters
        last4 = last4_quarters_map.get(pd.Timestamp(date_ref))
        if last4:
            cfo_sum = sum([cfo_map.get(q) or 0 for q in last4]) if all([cfo_map.get(q) is not None for q in last4]) else None
            cap_sum = sum([capex_map.get(q) or 0 for q in last4]) if all([capex_map.get(q) is not None for q in last4]) else None
            if cfo_sum is not None and cap_sum is not None:
                fcf_ttm = cfo_sum - cap_sum
        _box(row_fcf_ttm, "FCF TTM ($m)", (fcf_ttm / 1e6) if fcf_ttm is not None else None, False, "#,##0.000")
        adj_fcf_ttm_val = adj_fcf_ttm_map.get(date_ref) if adj_fcf_ttm_map else None
        _box(row_adj_fcf_ttm, "Adj FCF TTM ($m)", (adj_fcf_ttm_val / 1e6) if adj_fcf_ttm_val is not None else None, False, "#,##0.000")
        rev_ttm_val = rev_ttm_map.get(date_ref) if 'rev_ttm_map' in locals() else None
        _box(row_rev_ttm, "Revenue TTM ($m)", (rev_ttm_val / 1e6) if rev_ttm_val is not None else None, False, "#,##0.000")
        capex_ttm_val = capex_ttm_map.get(date_ref)
        _box(row_capex_ttm, "Capex TTM ($m)", (capex_ttm_val / 1e6) if capex_ttm_val is not None else None, False, "#,##0.000")
        int_paid_ttm_val = int_paid_ttm_map.get(date_ref) if int_paid_ttm_map else None
        _box(row_int_paid_ttm, "Interest paid TTM ($m)", (int_paid_ttm_val / 1e6) if int_paid_ttm_val is not None else None, False, "#,##0.000")
        # EPS TTM
        eps_ttm = None
        if last4:
            ni_vals = [net_income_map.get(q) for q in last4]
            sh_vals = [shares_map.get(q) for q in last4]
            if all(v is not None for v in ni_vals) and all(v is not None for v in sh_vals) and all(v != 0 for v in sh_vals):
                ni_ttm = float(sum(ni_vals))
                sh_avg = float(sum(sh_vals)) / 4.0
                eps_ttm = ni_ttm / sh_avg if sh_avg != 0 else None
        _box(row_eps_ttm, "EPS TTM ($)", eps_ttm, False, "#,##0.000")
        adj_eps_ttm_val = adj_eps_ttm_map.get(date_ref) if adj_eps_ttm_map else None
        _box(row_adj_eps_ttm, "Adj EPS TTM ($)", adj_eps_ttm_val, False, "#,##0.000")
        _box(row_bv, "BV/share", bv_share_map.get(date_ref), False, "#,##0.000")
        _box(row_tbv, "TBV/share", tbv_share_map.get(date_ref), False, "#,##0.000")
    _box(row_tgt_ev_adj, "Target EV/Adj EBITDA", 6.0, True, "0.0x")
    _box(row_tgt_ev, "Target EV/EBITDA", 6.0, True, "0.0x")
    _box(row_tgt_fcf, "Target EV yield", 0.10, True, "0.0%")
    _box(row_owner_maint_ratio, "Maint. capex % of capex", owner_maint_capex_ratio_default, True, "0.0%")
    _box(row_owner_recurring, "Recurring cash costs ($m)", 0.0, True, "#,##0.000")
    _box(row_owner_wc_norm, "WC normalization ($m)", 0.0, True, "#,##0.000")
    _box(row_share_mode, "Per-share denominator", "Diluted", True)

    input_basis_hint_map: Dict[int, Tuple[str, str]] = {
        row_price: ("User", "Primary market input. Update when testing valuation scenarios."),
        row_asof: ("Latest Q", "Reference quarter used for TTM and current balance-sheet metrics."),
        row_shares_out: ("Latest Q", "Used for per-share outputs when mode = Outstanding."),
        row_shares_dil: ("Latest Q", "Used for per-share outputs when mode = Diluted."),
        row_net_debt: ("Latest Q", "Debt core minus cash at as-of quarter."),
        row_ebitda_ttm: ("TTM", "Base GAAP-like EBITDA over trailing 4 quarters."),
        row_adj_ebitda_ttm: ("TTM", "Adjusted EBITDA over trailing 4 quarters."),
        row_fcf_ttm: ("TTM", "FCF proxy from CFO - Capex over trailing 4 quarters."),
        row_adj_fcf_ttm: ("TTM", "Non-GAAP FCF when available; otherwise blank."),
        row_rev_ttm: ("TTM", "Revenue over trailing 4 quarters."),
        row_eps_ttm: ("TTM", "GAAP EPS over trailing 4 quarters."),
        row_adj_eps_ttm: ("TTM", "Adjusted EPS over trailing 4 quarters."),
        row_bv: ("Latest Q", "Book value per share from latest quarter balance sheet."),
        row_tbv: ("Latest Q", "Tangible book value per share from latest quarter."),
        row_tgt_ev_adj: ("User", "Target EV/Adj EBITDA multiple; adjust for business quality and cycle."),
        row_tgt_ev: ("User", "Target EV/EBITDA multiple used as cross-check."),
        row_tgt_fcf: ("User", "Target EV yield accepts 0.10 or 10%."),
        row_capex_ttm: ("TTM", "Capex trailing 4 quarters."),
        row_int_paid_ttm: ("TTM", "Cash interest paid trailing 4 quarters."),
        row_owner_maint_ratio: ("Assumption", "Maintenance share of capex used in owner-earnings bridge."),
        row_owner_recurring: ("Assumption", "Recurring cash costs to subtract from owner earnings."),
        row_owner_wc_norm: ("Assumption", "Working-capital normalization adjustment."),
        row_share_mode: ("User", "Choose Outstanding or Diluted denominator for per-share values."),
    }
    for _rr, (_basis, _hint) in input_basis_hint_map.items():
        _set_input_meta(_rr, _basis, _hint)

    named_inputs = {
        "Price": row_price,
        "AsOfQuarter": row_asof,
        "Shares": row_shares_out,
        "SharesDiluted": row_shares_dil,
        "NetDebt": row_net_debt,
        "Base_EBITDA": row_ebitda_ttm,
        "Adj_EBITDA": row_adj_ebitda_ttm,
        "FCF_TTM": row_fcf_ttm,
        "Adj_FCF_TTM": row_adj_fcf_ttm,
        "Revenue_TTM": row_rev_ttm,
        "EPS_TTM": row_eps_ttm,
        "Adj_EPS_TTM": row_adj_eps_ttm,
        "BV_PerShare": row_bv,
        "TBV_PerShare": row_tbv,
        "Target_EV_AdjEBITDA": row_tgt_ev_adj,
        "Target_EV_EBITDA": row_tgt_ev,
        "Target_EV_Yield": row_tgt_fcf,
        "Capex_TTM": row_capex_ttm,
        "InterestPaid_TTM": row_int_paid_ttm,
        "MaintCapexRatio": row_owner_maint_ratio,
        "RecurringCashCosts": row_owner_recurring,
        "WCNormalization": row_owner_wc_norm,
        "PerShareMode": row_share_mode,
    }
    for _n, _r in named_inputs.items():
        input_cell = ws.cell(row=_r, column=input_value_col)
        input_cell.value = _normalize_thesis_bridge_basis(_n, input_cell.value)
        _set_named_range(_n, _r, input_value_col)
    if quarter_columns:
        latest_visible_col = get_column_letter(quarter_columns[-1])
        _set_named_range_ref("CompanyOperatingMargin_Latest", f"'Valuation'!${latest_visible_col}${row_operating_margin_pct}")
        _set_named_range_ref("OperatingMargin_Latest", f"'Valuation'!${latest_visible_col}${row_operating_margin_pct}")
        _set_named_range_ref("CompanyOperatingMargin_TTM", f"'Valuation'!${latest_visible_col}${row_operating_margin_ttm_pct}")
    # Inline normalized target-yield expression (accepts both 0.10 and 10%/10).
    # Keep it inline in formulas to avoid helper cells / workbook defined-name issues.
    target_ev_yield_n_expr = 'IF(OR(Target_EV_Yield="",Target_EV_Yield<=0),"",IF(Target_EV_Yield>1,Target_EV_Yield/100,Target_EV_Yield))'

    # Outputs (named-range formulas; locale-safe)
    out_col = get_column_letter(output_value_col)
    in_col = get_column_letter(input_value_col)
    market_col = get_column_letter(market_value_col)
    scn_col = get_column_letter(scn_value_col)
    dcf_col = get_column_letter(dcf_value_col)
    output_interp_col_letter = get_column_letter(output_interp_col)
    market_interp_col_letter = get_column_letter(market_interp_col)
    _box(row_mktcap, "Market cap ($m)", "=IF(OR(Price=\"\",Shares=\"\"),\"\",Price*Shares)", False, "#,##0.000", output_label_col, output_value_col)
    _box(row_ev, "EV ($m)", "=IF(OR(MarketCap=\"\",NetDebt=\"\"),\"\",MarketCap+NetDebt)", False, "#,##0.000", output_label_col, output_value_col)
    _box(row_implied_ev_adj, "Implied EV/Adj EBITDA", "=IF(OR(EV=\"\",Adj_EBITDA=\"\",Adj_EBITDA<=0),\"\",EV/Adj_EBITDA)", False, "0.00x", output_label_col, output_value_col)
    _box(row_implied_ev, "Implied EV/EBITDA", "=IF(OR(EV=\"\",Base_EBITDA=\"\",Base_EBITDA<=0),\"\",EV/Base_EBITDA)", False, "0.00x", output_label_col, output_value_col)
    _box(
        row_fcff_proxy_ttm,
        "FCFF proxy TTM ($m)",
        "=IFERROR(IF(OR(IFERROR(FCF_TTM,\"\")=\"\",IFERROR(InterestPaid_TTM,\"\")=\"\"),\"\",N(FCF_TTM)+N(InterestPaid_TTM)),\"\")",
        False,
        "#,##0.000",
        output_label_col,
        output_value_col,
    )
    _box(row_implied_fcff, "Implied FCFF yield (EV)", "=IF(OR(EV=\"\",FCFF_Proxy_TTM=\"\",EV<=0),\"\",FCFF_Proxy_TTM/EV)", False, "0.0%", output_label_col, output_value_col)
    _box(row_equity_fcf, "Equity FCF yield", "=IF(OR(MarketCap=\"\",FCF_TTM=\"\",MarketCap<=0),\"\",FCF_TTM/MarketCap)", False, "0.0%", output_label_col, output_value_col)
    _box(
        row_owner_fcf_ttm,
        "Owner earnings TTM ($m)",
        "=IF(OR(FCF_TTM=\"\",Capex_TTM=\"\",MaintCapexRatio=\"\"),\"\",FCF_TTM+(1-MaintCapexRatio)*Capex_TTM-RecurringCashCosts+WCNormalization)",
        False,
        "#,##0.000",
        output_label_col,
        output_value_col,
    )
    _box(row_owner_fcf_yield, "Owner earnings yield", "=IF(OR(EV=\"\",OwnerEarnings_TTM=\"\",EV<=0),\"\",OwnerEarnings_TTM/EV)", False, "0.0%", output_label_col, output_value_col)
    fair_denom = "IF(PerShareMode=\"Outstanding\",Shares,SharesDiluted)"
    _box(row_eq_adj, "Eq/Share @ target EV/Adj EBITDA", f"=IF(OR({fair_denom}=\"\",Adj_EBITDA=\"\"),\"\",(Target_EV_AdjEBITDA*Adj_EBITDA-NetDebt)/{fair_denom})", False, "$#,##0.00", output_label_col, output_value_col)
    _box(row_eq_ev, "Eq/Share @ target EV/EBITDA", f"=IF(OR({fair_denom}=\"\",Base_EBITDA=\"\"),\"\",(Target_EV_EBITDA*Base_EBITDA-NetDebt)/{fair_denom})", False, "$#,##0.00", output_label_col, output_value_col)
    _box(
        row_eq_fcf,
        "Eq/Share @ target EV yield (FCFF)",
        f"=IF(OR({fair_denom}=\"\",FCFF_Proxy_TTM=\"\",FCFF_Proxy_TTM<=0,({target_ev_yield_n_expr})=\"\",({target_ev_yield_n_expr})<=0),\"\",((FCFF_Proxy_TTM/({target_ev_yield_n_expr}))-NetDebt)/{fair_denom})",
        False,
        "$#,##0.00",
        output_label_col,
        output_value_col,
    )
    _box(row_pe, "P/E (TTM)", "=IF(EPS_TTM=\"\",\"\",IF(EPS_TTM<=0,\"\",Price/EPS_TTM))", False, "0.00x", output_label_col, output_value_col)
    _box(row_pe_adj, "P/E (Adj TTM)", "=IF(Adj_EPS_TTM=\"\",\"\",IF(Adj_EPS_TTM<=0,\"\",Price/Adj_EPS_TTM))", False, "0.00x", output_label_col, output_value_col)
    _box(row_ev_sales, "EV/Sales (TTM)", "=IF(OR(EV=\"\",Revenue_TTM=\"\",Revenue_TTM<=0),\"\",EV/Revenue_TTM)", False, "0.00x", output_label_col, output_value_col)
    _box(row_pb, "Price/BV", "=IF(OR(Price=\"\",Price<=0,BV_PerShare=\"\"),\"\",IF(BV_PerShare<=0,\"n/a (neg equity)\",Price/BV_PerShare))", False, None, output_label_col, output_value_col)
    _box(row_ptbv, "Price/TBV", "=IF(OR(Price=\"\",Price<=0,TBV_PerShare=\"\"),\"\",IF(TBV_PerShare<=0,\"n/a (neg equity)\",Price/TBV_PerShare))", False, None, output_label_col, output_value_col)

    # Output interpretations (display-only, deterministic).
    _set_interpretation(row_mktcap, "Price × shares; equity market value implied by current input price.", output_interp_col)
    _set_interpretation(row_ev, "Market cap plus net debt; enterprise value used across multiples/yields.", output_interp_col)
    _set_interpretation(row_implied_ev_adj, "Current price implies this EV/Adj EBITDA multiple; compare vs target multiple.", output_interp_col)
    _set_interpretation(row_implied_ev, "Current price implies this EV/EBITDA multiple; cross-check vs target EV/EBITDA.", output_interp_col)
    _set_interpretation(row_fcff_proxy_ttm, "FCF + cash interest paid; proxy FCFF base used for target-yield math.", output_interp_col)
    _set_interpretation(row_implied_fcff, "Current EV translated to FCFF yield proxy.", output_interp_col)
    _set_interpretation(row_equity_fcf, "FCF relative to equity value only (market cap).", output_interp_col)
    _set_interpretation(row_owner_fcf_ttm, "Owner earnings bridge after maintenance capex and recurring adjustments.", output_interp_col)
    _set_interpretation(row_owner_fcf_yield, "Owner earnings relative to enterprise value.", output_interp_col)
    _set_interpretation(row_eq_adj, "Per-share value if target EV/Adj EBITDA holds today.", output_interp_col)
    _set_interpretation(row_eq_ev, "Per-share value if target EV/EBITDA holds today.", output_interp_col)
    _set_interpretation(row_eq_fcf, "Per-share value using FCFF target yield framework.", output_interp_col)
    _set_interpretation(row_pe, "Price divided by GAAP EPS TTM.", output_interp_col)
    _set_interpretation(row_pe_adj, "Price divided by adjusted EPS TTM.", output_interp_col)
    _set_interpretation(row_ev_sales, "Enterprise value divided by revenue TTM.", output_interp_col)
    _set_interpretation(row_pb, "Price to book value per share; n/a when equity is negative.", output_interp_col)
    _set_interpretation(row_ptbv, "Price to tangible book value per share; n/a when TBV is negative.", output_interp_col)
    if is_anf_profile and not valuation_price_input_available:
        # ANF has no market-data source enabled. Keep scenario/target rows, but
        # do not show live price-linked formulas that render as #NAME? in the
        # desktop preview when the user has not supplied a price.
        for _rr in (
            row_mktcap,
            row_ev,
            row_implied_ev_adj,
            row_implied_ev,
            row_implied_fcff,
            row_equity_fcf,
            row_owner_fcf_yield,
            row_pe,
            row_pe_adj,
            row_ev_sales,
            row_pb,
            row_ptbv,
        ):
            ws.cell(row=_rr, column=output_value_col).value = None

    named_outputs = {
        "MarketCap": row_mktcap,
        "EV": row_ev,
        "Implied_EV_AdjEBITDA": row_implied_ev_adj,
        "Implied_EV_EBITDA": row_implied_ev,
        "FCFF_Proxy_TTM": row_fcff_proxy_ttm,
        "Implied_FCFF_Yield": row_implied_fcff,
        "Equity_FCF_Yield": row_equity_fcf,
        "OwnerEarnings_TTM": row_owner_fcf_ttm,
        "OwnerEarnings_Yield": row_owner_fcf_yield,
        "EqShare_Target_Adj": row_eq_adj,
        "EqShare_Target_EV": row_eq_ev,
        "EqShare_Target_Yield": row_eq_fcf,
    }
    for _n, _r in named_outputs.items():
        _set_named_range(_n, _r, output_value_col)

    qadj_col = get_column_letter(qadj_value_col)
    ws.cell(row=row_market_hdr, column=market_label_col, value="What Market Is Pricing").font = bold
    ws.cell(row=row_market_hdr, column=market_label_col).fill = section_fill
    try:
        ws.merge_cells(
            start_row=row_market_hdr,
            start_column=market_label_col,
            end_row=row_market_hdr,
            end_column=market_interp_col + 4,
        )  # L:U
    except Exception:
        pass
    _box(
        row_req_adj_ebitda,
        "Required Adj EBITDA @ target multiple ($m)",
        f"=IF(OR(EV=\"\",{qadj_col}{row_qadj_ev_adj}=\"\",{qadj_col}{row_qadj_ev_adj}<=0),\"\",EV/{qadj_col}{row_qadj_ev_adj})",
        False,
        "#,##0.000",
        market_label_col,
        market_value_col,
    )
    _box(
        row_req_adj_delta,
        "Implied Adj EBITDA change vs current",
        f"=IFERROR(IF(AND({market_col}{row_req_adj_ebitda}>0,Adj_EBITDA>0),{market_col}{row_req_adj_ebitda}/Adj_EBITDA-1,\"\"),\"\")",
        False,
        "0.0%",
        market_label_col,
        market_value_col,
    )
    _box(
        row_req_fcff,
        "Required FCFF @ target EV yield ($m)",
        f"=IF(OR(EV=\"\",{qadj_col}{row_qadj_yield}=\"\",{qadj_col}{row_qadj_yield}<=0),\"\",EV*{qadj_col}{row_qadj_yield})",
        False,
        "#,##0.000",
        market_label_col,
        market_value_col,
    )
    _box(
        row_req_fcff_delta,
        "Implied FCFF change vs current",
        f"=IFERROR(IF(AND(IFERROR(N({market_col}{row_req_fcff}),0)>0,IFERROR(N(FCFF_Proxy_TTM),0)>0),N({market_col}{row_req_fcff})/IFERROR(N(FCFF_Proxy_TTM),0)-1,\"\"),\"\")",
        False,
        "0.0%",
        market_label_col,
        market_value_col,
    )
    _box(
        row_req_owner_fcf,
        "Required owner earnings @ target EV yield ($m)",
        f"=IF(OR(EV=\"\",{qadj_col}{row_qadj_yield}=\"\",{qadj_col}{row_qadj_yield}<=0),\"\",EV*{qadj_col}{row_qadj_yield})",
        False,
        "#,##0.000",
        market_label_col,
        market_value_col,
    )
    _box(
        row_req_owner_delta,
        "Implied owner earnings change",
        f"=IFERROR(IF(AND({market_col}{row_req_owner_fcf}>0,OwnerEarnings_TTM>0),{market_col}{row_req_owner_fcf}/OwnerEarnings_TTM-1,\"\"),\"\")",
        False,
        "0.0%",
        market_label_col,
        market_value_col,
    )
    _set_interpretation(row_req_adj_ebitda, "At quality-adjusted target multiple, Adj EBITDA must reach this level.", market_interp_col)
    _set_interpretation(row_req_adj_delta, "Required Adj EBITDA change versus current TTM level.", market_interp_col)
    _set_interpretation(row_req_fcff, "At quality-adjusted target yield, FCFF requirement implied by EV.", market_interp_col)
    _set_interpretation(row_req_fcff_delta, "Market-implied FCFF delta versus FCFF proxy TTM.", market_interp_col)
    _set_interpretation(row_req_owner_fcf, "Owner earnings required at quality-adjusted target yield.", market_interp_col)
    _set_interpretation(row_req_owner_delta, "Owner earnings change versus current owner-earnings TTM.", market_interp_col)

    # Market-implied terminal g block (solve gT so DCF EV = Market EV).
    ws.cell(row=row_mi_hdr, column=16, value="Market-implied terminal g (solve gT so DCF EV = Market EV)").font = bold
    ws.cell(row=row_mi_hdr, column=16).fill = section_fill
    try:
        ws.merge_cells(start_row=row_mi_hdr, start_column=16, end_row=row_mi_hdr, end_column=19)  # P:S
    except Exception:
        pass
    ws.cell(row=row_mi_market_ev, column=16, value="Market EV ($m)")
    ws.cell(row=row_mi_market_ev, column=17, value="=EV").number_format = "#,##0.000"
    ws.cell(row=row_mi_dcf_ev, column=16, value="DCF EV ($m)")
    ws.cell(row=row_mi_dcf_ev, column=17, value=f"={dcf_col}{row_dcf_ev}").number_format = "#,##0.000"
    ws.cell(row=row_mi_curr_wacc, column=16, value="Current WACC (cell)")
    ws.cell(row=row_mi_curr_wacc, column=17, value=f"={dcf_col}{row_dcf_wacc}").number_format = "0.0%"
    ws.cell(row=row_mi_curr_gt, column=16, value="Current gT (cell)")
    ws.cell(row=row_mi_curr_gt, column=17, value=f"={dcf_col}{row_dcf_gt}").number_format = "0.0%"
    ws.cell(row=row_mi_tbl_hdr, column=16, value="WACC").font = bold
    ws.cell(row=row_mi_tbl_hdr, column=17, value="Implied terminal g").font = bold
    try:
        ws.merge_cells(start_row=row_mi_tbl_hdr, start_column=17, end_row=row_mi_tbl_hdr, end_column=18)  # Q:R
    except Exception:
        pass
    ws.cell(row=row_mi_tbl_hdr, column=19, value="Status").font = bold
    for cc in (16, 17, 18, 19):
        ws.cell(row=row_mi_tbl_hdr, column=cc).fill = header_fill
    for i, wacc_v in enumerate(dcf_sens_wacc_vals):
        rr = row_mi_wacc_start + i
        ws.cell(row=rr, column=16, value=wacc_v).number_format = "0.0%"
        ws.cell(row=rr, column=17).number_format = "0.0%"
        ws.cell(row=rr, column=19, value="").alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    try:
        ws.merge_cells(start_row=row_mi_toggle, start_column=16, end_row=row_mi_toggle, end_column=17)
    except Exception:
        pass
    ws.cell(row=row_mi_toggle, column=16, value="Auto-update implied gT")
    t_cell = ws.cell(row=row_mi_toggle, column=18, value=True)
    t_cell.fill = input_fill
    t_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    _set_named_range("AutoImpliedGT", row_mi_toggle, 18)
    _set_named_range_ref("ImpliedGT_WACC", f"'Valuation'!$P${row_mi_wacc_start}:$P${row_mi_wacc_end}")
    _set_named_range_ref("ImpliedGT_Output", f"'Valuation'!$Q${row_mi_wacc_start}:$Q${row_mi_wacc_end}")
    _set_named_range_ref("ImpliedGT_Status", f"'Valuation'!$S${row_mi_wacc_start}:$S${row_mi_wacc_end}")
    try:
        ws.conditional_formatting.add(
            f"Q{row_mi_wacc_start}:Q{row_mi_wacc_end}",
            CellIsRule(operator="greaterThan", formula=["0.035"], fill=PatternFill("solid", fgColor="F4B183")),
        )
        ws.conditional_formatting.add(
            f"Q{row_mi_wacc_start}:Q{row_mi_wacc_end}",
            CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="F8CBAD")),
        )
    except Exception:
        pass
    if is_anf_profile and not valuation_price_input_available:
        # Keep the PBI-style market-pricing framework visible, but do not
        # display name-reference formulas that require a user-supplied price.
        for _rr in (
            row_req_adj_ebitda,
            row_req_adj_delta,
            row_req_fcff,
            row_req_fcff_delta,
            row_req_owner_fcf,
            row_req_owner_delta,
        ):
            ws.cell(row=_rr, column=market_value_col).value = None
        ws.cell(row=row_mi_market_ev, column=17).value = None
        for _rr in range(row_mi_wacc_start, row_mi_wacc_end + 1):
            ws.cell(row=_rr, column=17).value = None
            ws.cell(row=_rr, column=19).value = ""

    # Scenario defaults from recent history (latest 8 available points).
    hist_growth: List[float] = []
    hist_margin: List[float] = []
    q_hist = [pd.Timestamp(q) for q in qs] if qs else []
    for i_q, qv in enumerate(q_hist):
        rev_now = rev_map.get(qv)
        if rev_now not in (None, 0):
            adj_now = (adj_ebitda_map.get(qv) if adj_ebitda_map else None)
            if adj_now is not None and pd.notna(adj_now):
                hist_margin.append(float(adj_now) / float(rev_now))
        if i_q >= 4:
            rev_prev = rev_map.get(q_hist[i_q - 4])
            if rev_now not in (None, 0) and rev_prev not in (None, 0):
                hist_growth.append(float(rev_now) / float(rev_prev) - 1.0)
    hist_growth = hist_growth[-8:]
    hist_margin = hist_margin[-8:]

    def _pct_quant(vals: List[float], qv: float, fallback: float) -> float:
        if not vals:
            return float(fallback)
        s = pd.Series(vals, dtype="float64").dropna()
        if s.empty:
            return float(fallback)
        try:
            return float(s.quantile(qv))
        except Exception:
            return float(fallback)

    base_growth_d = _pct_quant(hist_growth, 0.50, 0.00)
    bull_growth_d = _pct_quant(hist_growth, 0.85, 0.02)
    bear_growth_d = _pct_quant(hist_growth, 0.15, -0.03)
    base_margin_d = _pct_quant(hist_margin, 0.50, 0.20)
    bull_margin_d = _pct_quant(hist_margin, 0.85, min(0.35, base_margin_d + 0.01))
    bear_margin_d = _pct_quant(hist_margin, 0.15, max(0.05, base_margin_d - 0.01))

    ws.cell(row=row_scn_hdr, column=scn_label_col, value="Trigger Scenarios").font = bold
    ws.cell(row=row_scn_hdr, column=scn_label_col).fill = section_fill
    try:
        ws.merge_cells(
            start_row=row_scn_hdr,
            start_column=scn_label_col,
            end_row=row_scn_hdr,
            end_column=scn_interp_col + 5,
        )
    except Exception:
        pass
    _box(row_scn_profile, "Scenario preset (Base/Bull/Bear/Custom)", "Base", True, None, scn_label_col, scn_value_col)
    _box(
        row_scn_growth,
        "Rev growth assumption",
        f"=IF({scn_col}{row_scn_profile}=\"Bull\",{bull_growth_d:.6f},IF({scn_col}{row_scn_profile}=\"Bear\",{bear_growth_d:.6f},IF({scn_col}{row_scn_profile}=\"Base\",{base_growth_d:.6f},{base_growth_d:.6f})))",
        True,
        "0.0%",
        scn_label_col,
        scn_value_col,
    )
    _box(
        row_scn_margin,
        "Adj EBITDA margin assumption",
        f"=IF({scn_col}{row_scn_profile}=\"Bull\",{bull_margin_d:.6f},IF({scn_col}{row_scn_profile}=\"Bear\",{bear_margin_d:.6f},IF({scn_col}{row_scn_profile}=\"Base\",{base_margin_d:.6f},{base_margin_d:.6f})))",
        True,
        "0.0%",
        scn_label_col,
        scn_value_col,
    )
    _box(row_scn_refi, "Refi interest normalization ($m)", f"=IF({scn_col}{row_scn_profile}=\"Bull\",15,IF({scn_col}{row_scn_profile}=\"Bear\",-10,0))", True, "#,##0.000", scn_label_col, scn_value_col)
    _box(row_scn_buyback, "Buyback continuation (m shares)", f"=IF({scn_col}{row_scn_profile}=\"Bull\",2,IF({scn_col}{row_scn_profile}=\"Bear\",-1,0))", True, "#,##0.000", scn_label_col, scn_value_col)
    _box(
        row_scn_adj_ebitda,
        "Scenario Adj EBITDA ($m)",
        _wrap_iferror(
            f"=IF(OR(Revenue_TTM=\"\",{scn_col}{row_scn_growth}=\"\",{scn_col}{row_scn_margin}=\"\"),\"\",Revenue_TTM*(1+{scn_col}{row_scn_growth})*{scn_col}{row_scn_margin})"
        ),
        False,
        "#,##0.000",
        scn_label_col,
        scn_value_col,
    )
    _box(
        row_scn_owner_fcf,
        "Scenario owner earnings ($m)",
        _wrap_iferror(
            f"=IF(OR(OwnerEarnings_TTM=\"\",{scn_col}{row_scn_adj_ebitda}=\"\",Adj_EBITDA=\"\"),\"\",OwnerEarnings_TTM+({scn_col}{row_scn_adj_ebitda}-Adj_EBITDA)+{scn_col}{row_scn_refi})"
        ),
        False,
        "#,##0.000",
        scn_label_col,
        scn_value_col,
    )
    _box(
        row_scn_eq_ev,
        "Eq/share scenario (EV/Adj)",
        _wrap_iferror(
            f"=IF(OR({scn_col}{row_scn_adj_ebitda}=\"\",{fair_denom}=\"\",{qadj_col}{row_qadj_ev_adj}=\"\",{qadj_col}{row_qadj_ev_adj}<=0),\"\",(({qadj_col}{row_qadj_ev_adj}*{scn_col}{row_scn_adj_ebitda})-NetDebt)/MAX(0.001,{fair_denom}-{scn_col}{row_scn_buyback}))"
        ),
        False,
        "$#,##0.00",
        scn_label_col,
        scn_value_col,
    )
    _box(
        row_scn_eq_fcf,
        "Eq/share scenario (EV owner earnings yield)",
        _wrap_iferror(
            f"=IF(OR({scn_col}{row_scn_owner_fcf}=\"\",{scn_col}{row_scn_owner_fcf}<=0,{fair_denom}=\"\",{qadj_col}{row_qadj_yield}=\"\",{qadj_col}{row_qadj_yield}<=0),\"\",(({scn_col}{row_scn_owner_fcf}/{qadj_col}{row_qadj_yield})-NetDebt)/MAX(0.001,{fair_denom}-{scn_col}{row_scn_buyback}))"
        ),
        False,
        "$#,##0.00",
        scn_label_col,
        scn_value_col,
    )
    c_profile_interp = ws.cell(
        row=row_scn_profile,
        column=scn_interp_col,
        value=_wrap_iferror(
            f"=IF(OR({scn_col}{row_scn_profile}=\"Base\",{scn_col}{row_scn_profile}=\"Bull\",{scn_col}{row_scn_profile}=\"Bear\"),\"Preset active\",\"Custom\")"
        ),
    )
    c_profile_interp.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c_profile_interp.font = Font(size=max(8, font_size - 1))
    _set_cell_comment(
        ws.cell(row=row_scn_profile, column=scn_value_col),
        "Preset fills defaults; overrides allowed on yellow scenario cells.",
    )
    _set_interpretation(row_scn_growth, "Default from historical revenue growth distribution (recent 8 quarters).", scn_interp_col)
    _set_interpretation(row_scn_margin, "Default from historical adjusted margin distribution (recent 8 quarters).", scn_interp_col)
    _set_interpretation(row_scn_refi, "Manual refinancing normalization assumption.", scn_interp_col)
    _set_interpretation(row_scn_buyback, "Manual share-count continuation assumption.", scn_interp_col)
    _set_interpretation(row_scn_adj_ebitda, "Scenario revenue × scenario margin.", scn_interp_col)
    _set_interpretation(row_scn_owner_fcf, "Owner earnings bridge adjusted by scenario EBITDA and refinancing.", scn_interp_col)
    _set_interpretation(row_scn_eq_ev, "Scenario per-share value with quality-adjusted EV/Adj multiple.", scn_interp_col)
    _set_interpretation(row_scn_eq_fcf, "Scenario per-share value with quality-adjusted owner-earnings yield.", scn_interp_col)

    for _n, _r in {
        "ScenarioProfile": row_scn_profile,
        "ScenarioGrowth": row_scn_growth,
        "ScenarioMargin": row_scn_margin,
        "ScenarioRefiNorm": row_scn_refi,
        "ScenarioBuyback": row_scn_buyback,
        "ScenarioAdjEBITDA": row_scn_adj_ebitda,
        "ScenarioOwnerEarnings": row_scn_owner_fcf,
        "ScenarioEqShare_EVAdj": row_scn_eq_ev,
        "ScenarioEqShare_Yield": row_scn_eq_fcf,
    }.items():
        _set_named_range(_n, _r, scn_value_col)
    try:
        ws.row_dimensions[row_drv_hdr - 1].hidden = False
        ws.row_dimensions[row_drv_hdr - 1].height = 18.0
    except Exception:
        pass
    try:
        for cc in range(scn_label_col, scn_interp_col + 6):
            _c_sep = ws.cell(row=row_drv_hdr, column=cc)
            _c_sep.border = Border(
                left=_c_sep.border.left,
                right=_c_sep.border.right,
                top=Side(style="thick", color="000000"),
                bottom=_c_sep.border.bottom,
            )
    except Exception:
        pass

    # Compact internal scenario drivers + quality toggles (portable, no external calls).
    def _fmt_source_comment(src: Dict[str, Any]) -> str:
        if not src:
            return "Source: N/A"
        filed = pd.to_datetime(src.get("filed"), errors="coerce")
        filed_txt = filed.strftime("%Y-%m-%d") if pd.notna(filed) else ""
        bits = [
            f"source={src.get('source_type') or src.get('method') or 'internal'}",
            f"form={src.get('form') or ''}",
            f"accn={src.get('accn') or ''}",
            f"doc={src.get('doc') or src.get('doc_path') or ''}",
            f"filed={filed_txt}",
        ]
        return " | ".join([b for b in bits if b and not b.endswith("=")])

    def _pick_note_for_driver(
        *,
        bucket_terms: List[str],
        metric_terms: List[str],
    ) -> Tuple[str, str]:
        if quarter_notes is None or quarter_notes.empty or date_ref is None:
            return ("N/A", "Source: N/A (Quarter_Notes unavailable)")
        d = _quarter_notes_view(quarter_mode="date")
        txt_cols = [c for c in [_resolve_col(d, ["claim", "headline"]), _resolve_col(d, ["note", "body"]), _resolve_col(d, ["evidence_snippet", "snippet"])] if c]
        if not txt_cols:
            return ("N/A", "Source: N/A (text columns missing)")
        cat_col = _resolve_col(d, ["category", "tag", "topic"])
        metric_col_qn = _resolve_col(d, ["metric", "metric_ref", "metric_tag"])
        score_col_qn = _resolve_col(d, ["severity_score", "score"])
        d = d[d["_quarter"] == pd.Timestamp(date_ref).date()].copy()
        if d.empty:
            return ("N/A", "Source: N/A (no quarter_notes for latest quarter)")
        best: Optional[Tuple[float, str, str]] = None
        for _, rr in d.iterrows():
            parts = [str(rr.get(tc) or "") for tc in txt_cols]
            txt = re.sub(r"\s+", " ", " ".join([p for p in parts if p]).strip())
            if not txt:
                continue
            low = txt.lower()
            cat_txt = str(rr.get(cat_col) or "").lower() if cat_col else ""
            metric_txt = str(rr.get(metric_col_qn) or "").lower() if metric_col_qn else ""
            score = float(pd.to_numeric(rr.get(score_col_qn), errors="coerce")) if score_col_qn else 0.0
            metric_hits = sum(1 for t in metric_terms if t and t.lower() in low)
            if metric_hits <= 0:
                continue
            score += 6.0 * metric_hits
            score += 2.0 * sum(1 for t in bucket_terms if t and (t.lower() in cat_txt or t.lower() in low or t.lower() in metric_txt))
            if score <= 0:
                continue
            src = {
                "source_type": rr.get("source_type") or rr.get("method") or "quarter_notes",
                "form": rr.get("form"),
                "accn": rr.get("accn"),
                "doc": rr.get("doc") or rr.get("doc_path"),
                "filed": rr.get("filed"),
            }
            snip = qn_compact_snippet(txt, 220)
            cand = (score, snip, _fmt_source_comment(src))
            if best is None or cand[0] > best[0]:
                best = cand
        if best is None:
            return ("N/A", "Source: N/A (no matching note)")
        return (best[1], best[2])

    ws.cell(row=row_drv_hdr, column=driver_label_col, value="Scenario drivers (internal)").font = bold
    ws.cell(row=row_drv_hdr, column=driver_label_col).fill = section_fill
    try:
        ws.merge_cells(
            start_row=row_drv_hdr,
            start_column=driver_label_col,
            end_row=row_drv_hdr,
            end_column=driver_label_col + 1,
        )
    except Exception:
        pass

    rev_drv_txt, rev_drv_src = _pick_note_for_driver(
        bucket_terms=["guidance", "results", "driver"],
        metric_terms=["revenue", "pricing", "volume", "mix", "demand"],
    )
    mar_drv_txt, mar_drv_src = _pick_note_for_driver(
        bucket_terms=["results", "driver", "program"],
        metric_terms=["margin", "cost savings", "sg&a", "r&d", "restructuring"],
    )
    fcf_drv_txt, fcf_drv_src = _pick_note_for_driver(
        bucket_terms=["cash flow", "fcf", "capex"],
        metric_terms=["fcf", "free cash flow", "capex", "cfo", "conversion"],
    )
    lev_drv_txt, lev_drv_src = _pick_note_for_driver(
        bucket_terms=["debt", "liquidity", "program"],
        metric_terms=["leverage", "debt", "refinancing", "buyback", "dividend", "capital allocation"],
    )

    _box(row_drv_rev, "Revenue trend driver", rev_drv_txt, False, None, driver_label_col, driver_value_col)
    _box(row_drv_margin, "Margin driver", mar_drv_txt, False, None, driver_label_col, driver_value_col)
    _box(row_drv_fcf, "FCF conversion driver", fcf_drv_txt, False, None, driver_label_col, driver_value_col)
    _box(row_drv_lev, "Leverage/cap allocation driver", lev_drv_txt, False, None, driver_label_col, driver_value_col)
    _set_snippet_with_comment(
        ws.cell(row=row_drv_rev, column=driver_value_col),
        rev_drv_txt,
        max_chars=110,
        extra_comment=rev_drv_src,
        visible_text=_clean_visible_driver_text_local(rev_drv_txt, max_chars=110),
    )
    _set_snippet_with_comment(
        ws.cell(row=row_drv_margin, column=driver_value_col),
        mar_drv_txt,
        max_chars=110,
        extra_comment=mar_drv_src,
        visible_text=_clean_visible_driver_text_local(mar_drv_txt, max_chars=110),
    )
    _set_snippet_with_comment(
        ws.cell(row=row_drv_fcf, column=driver_value_col),
        fcf_drv_txt,
        max_chars=110,
        extra_comment=fcf_drv_src,
        visible_text=_clean_visible_driver_text_local(fcf_drv_txt, max_chars=110),
    )
    _set_snippet_with_comment(
        ws.cell(row=row_drv_lev, column=driver_value_col),
        lev_drv_txt,
        max_chars=110,
        extra_comment=lev_drv_src,
        visible_text=_clean_visible_driver_text_local(lev_drv_txt, max_chars=110),
    )

    ws.cell(row=row_toggle_hdr, column=toggle_label_col, value="Quality toggles").font = bold
    ws.cell(row=row_toggle_hdr, column=toggle_label_col).fill = section_fill
    try:
        ws.merge_cells(
            start_row=row_toggle_hdr,
            start_column=toggle_label_col,
            end_row=row_toggle_hdr,
            end_column=toggle_label_col + 1,
        )
    except Exception:
        pass

    q0_latest = pd.Timestamp(date_ref) if date_ref is not None else None
    adj_gap_ratio: Optional[float] = None
    lev_latest: Optional[float] = None
    if q0_latest is not None:
        gaap_ttm_latest = ebitda_ttm_map.get(q0_latest)
        adj_ttm_latest = adj_ebitda_ttm_map.get(q0_latest) if adj_ebitda_ttm_map else None
        if gaap_ttm_latest is not None and adj_ttm_latest not in (None, 0):
            adj_gap_ratio = abs(float(adj_ttm_latest) - float(gaap_ttm_latest)) / abs(float(adj_ttm_latest))
        lev_latest = net_lev_map.get(q0_latest) if net_lev_map else None

    note_blob = ""
    if quarter_notes is not None and not quarter_notes.empty:
        qn_tmp = _quarter_notes_view(quarter_mode="date")
        if date_ref is not None:
            qn_tmp = qn_tmp[qn_tmp["_quarter"] == pd.Timestamp(date_ref).date()]
        txt_candidates = []
        for cc in ["claim", "note", "body", "headline", "evidence_snippet"]:
            if cc in qn_tmp.columns:
                txt_candidates.extend([str(v) for v in qn_tmp[cc].dropna().astype(str).tolist()])
        note_blob = " ".join(txt_candidates).lower()

    dep_blob = " ".join([str(x) for x in (company_overview or {}).get("key_dependencies", [])]).lower()
    regulatory_dep = bool(re.search(r"\b(usps|postal|regulatory)\b", note_blob + " " + dep_blob))
    persistent_gap = bool(adj_gap_ratio is not None and adj_gap_ratio >= 0.20)
    constrained_optionality = bool(lev_latest is not None and float(lev_latest) >= 3.0)
    customer_conc = bool(re.search(r"\b(customer concentration|significant customer|major customer)\b", note_blob + " " + dep_blob))

    _box(row_toggle_reg, "Regulatory dependency?", "Yes" if regulatory_dep else "No", False, None, toggle_label_col, toggle_value_col)
    _box(row_toggle_gap, "Persistent non-GAAP gap?", "Yes" if persistent_gap else "No", False, None, toggle_label_col, toggle_value_col)
    _box(row_toggle_lev, "Balance-sheet optionality constrained?", "Yes" if constrained_optionality else "No", False, None, toggle_label_col, toggle_value_col)
    _box(row_toggle_conc, "Customer concentration disclosed?", "Yes" if customer_conc else "No", False, None, toggle_label_col, toggle_value_col)

    _set_cell_comment(
        ws.cell(row=row_toggle_reg, column=toggle_value_col),
        "Evidence from Quarter_Notes/company overview dependency lines (USPS/postal/regulatory keywords).",
    )
    _set_cell_comment(
        ws.cell(row=row_toggle_gap, column=toggle_value_col),
        f"Computed from Adj EBITDA vs GAAP EBITDA gap ratio at latest quarter: {adj_gap_ratio:.1%}." if adj_gap_ratio is not None else "N/A",
    )
    _set_cell_comment(
        ws.cell(row=row_toggle_lev, column=toggle_value_col),
        f"Computed from corporate_net_leverage latest: {lev_latest:.2f}x." if lev_latest is not None else "N/A",
    )
    _set_cell_comment(
        ws.cell(row=row_toggle_conc, column=toggle_value_col),
        "Evidence from filing-derived dependency text (customer concentration keywords).",
    )

    ws.cell(row=row_qadj_hdr, column=qadj_label_col, value="Quality-adjusted target multiple/yield").font = bold
    ws.cell(row=row_qadj_hdr, column=qadj_label_col).fill = section_fill
    try:
        ws.merge_cells(
            start_row=row_qadj_hdr,
            start_column=qadj_label_col,
            end_row=row_qadj_hdr,
            end_column=qadj_label_col + 1,
        )
    except Exception:
        pass
    _box(
        row_qadj_ev_adj,
        "Adj target EV/Adj EBITDA (x)",
        f"=IF(Target_EV_AdjEBITDA=\"\",\"\",MAX(0,Target_EV_AdjEBITDA-IF({get_column_letter(toggle_value_col)}{row_toggle_lev}=\"Yes\",0.25,0)))",
        False,
        "0.00x",
        qadj_label_col,
        qadj_value_col,
    )
    _box(
        row_qadj_ev,
        "Adj target EV/EBITDA (x)",
        f"=IF(Target_EV_EBITDA=\"\",\"\",MAX(0,Target_EV_EBITDA-IF({get_column_letter(toggle_value_col)}{row_toggle_lev}=\"Yes\",0.25,0)))",
        False,
        "0.00x",
        qadj_label_col,
        qadj_value_col,
    )
    _box(
        row_qadj_yield,
        "Adj target EV yield",
        f"=IF(OR(({target_ev_yield_n_expr})=\"\",({target_ev_yield_n_expr})<=0),\"\",({target_ev_yield_n_expr})+IF({get_column_letter(toggle_value_col)}{row_toggle_reg}=\"Yes\",0.005,0)+IF({get_column_letter(toggle_value_col)}{row_toggle_gap}=\"Yes\",0.005,0)+IF({get_column_letter(toggle_value_col)}{row_toggle_conc}=\"Yes\",0.0025,0))",
        False,
        "0.0%",
        qadj_label_col,
        qadj_value_col,
    )
    _set_interpretation(row_qadj_ev_adj, "Multiple adjusted for balance-sheet optionality toggle.", qadj_text_col)
    _set_interpretation(row_qadj_ev, "Cross-check multiple adjusted for balance-sheet optionality toggle.", qadj_text_col)
    _set_interpretation(row_qadj_yield, "Yield uplift from regulatory/non-GAAP/customer-concentration toggles.", qadj_text_col)
    _set_cell_comment(
        ws.cell(row=row_qadj_yield, column=qadj_value_col),
        "Adjustments: regulatory +0.5pp, persistent non-GAAP gap +0.5pp, customer concentration +0.25pp.",
    )

    # Convertible optionality block.
    row_convert_hdr = 246
    convert_col_map = {
        "Security": 12,   # L:M
        "Principal": 14,  # N
        "Coupon": 15,  # O
        "Maturity": 16,  # P
        "Conversion price": 17,  # Q:R
        "Shares on full conversion (m)": 19,  # S:T
        "Concurrent repurchase shares (m)": 21,  # U:W
        "Net added shares on full conversion (m)": 24,  # X:AA
    }
    convert_header_end_col = 27
    net_added_end_col = 27
    ws.cell(row=row_convert_hdr, column=12, value="Convertible notes").font = bold
    ws.cell(row=row_convert_hdr, column=12).fill = section_fill
    try:
        ws.merge_cells(start_row=row_convert_hdr, start_column=12, end_row=row_convert_hdr, end_column=convert_header_end_col)
    except Exception:
        pass
    convert_header_row = row_convert_hdr + 1
    for hdr_txt, hdr_col in convert_col_map.items():
        c_hdr = ws.cell(row=convert_header_row, column=hdr_col, value=hdr_txt)
        c_hdr.font = bold
        c_hdr.fill = header_fill
        c_hdr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c_hdr.border = thin_border
    for merge_start, merge_end in ((12, 13), (17, 18), (19, 20), (21, 23), (24, net_added_end_col)):
        try:
            ws.merge_cells(start_row=convert_header_row, start_column=merge_start, end_row=convert_header_row, end_column=merge_end)
        except Exception:
            pass
    convert_end_col = 27
    conv_price_col = 17
    shares_full_col = 19
    concurrent_rep_col = 21
    net_added_col = 24
    ws.column_dimensions["L"].width = 13.0
    ws.column_dimensions["M"].width = 13.0
    shares_full_letter = get_column_letter(shares_full_col)
    concurrent_rep_letter = get_column_letter(concurrent_rep_col)

    def _usable_convertible_rows(df_in: Optional[pd.DataFrame]) -> pd.DataFrame:
        """Filter out QA sentinel rows before deciding whether fallback data is needed."""
        if df_in is None or getattr(df_in, "empty", True):
            return pd.DataFrame()
        df_local = df_in.copy()
        name_txt = df_local.get("tranche_name", pd.Series("", index=df_local.index)).astype(str)
        security_txt = df_local.get("security", pd.Series("", index=df_local.index)).astype(str)
        instr_txt = df_local.get("instrument_type", pd.Series("", index=df_local.index)).astype(str)
        review_txt = df_local.get("review_note", pd.Series("", index=df_local.index)).astype(str)
        qa_txt = df_local.get("qa_status", pd.Series("", index=df_local.index)).astype(str)
        combined_txt = (name_txt + " " + security_txt + " " + instr_txt + " " + review_txt + " " + qa_txt).str.lower()
        principal_num = pd.to_numeric(
            df_local.get("amount_principal", pd.Series(float("nan"), index=df_local.index)),
            errors="coerce",
        )
        sentinel_mask = combined_txt.str.contains(r"needs review|tie[- ]out failed|qa_guardrail", na=False)
        has_principal = principal_num.notna() & (principal_num > 0)
        convert_mask = (
            instr_txt.str.lower().eq("convertible")
            | combined_txt.str.contains(r"\bconvert(?:ible|ed|ion)?\b", na=False)
            | pd.to_numeric(df_local.get("conversion_price", pd.Series(float("nan"), index=df_local.index)), errors="coerce").notna()
            | pd.to_numeric(df_local.get("shares_on_full_conversion", pd.Series(float("nan"), index=df_local.index)), errors="coerce").notna()
        )
        return df_local[~sentinel_mask & has_principal & convert_mask].copy()

    convert_df = pd.DataFrame()
    if debt_tranches_latest is not None and not debt_tranches_latest.empty:
        convert_df = debt_tranches_latest.copy()
        if "instrument_type" in convert_df.columns:
            convert_df = convert_df[convert_df["instrument_type"].astype(str).str.lower().eq("convertible")].copy()
        convert_df = _usable_convertible_rows(convert_df)
    if (convert_df is None or convert_df.empty) and str(ticker).upper() == "GPRE":
        try:
            fallback_convert_df = _source_backed_debt_tranches_from_slides(
                slides_debt,
                qs[-1] if qs else None,
                ticker,
            )
        except Exception:
            fallback_convert_df = pd.DataFrame()
        if fallback_convert_df is not None and not fallback_convert_df.empty:
            name_series = fallback_convert_df.get("tranche_name", pd.Series(dtype=object)).astype(str)
            convert_df = fallback_convert_df[name_series.str.contains(r"\bconvertible\b", case=False, na=False)].copy()
            if not convert_df.empty and "instrument_type" not in convert_df.columns:
                convert_df["instrument_type"] = "convertible"
            convert_df = _usable_convertible_rows(convert_df)
    convert_row = convert_header_row + 1
    base_share_formula = 'IF(SharesDiluted<>"",SharesDiluted,Shares)'
    def _convertible_text_matches_row(txt_in: Any, row_in: pd.Series, require_convert: bool = True) -> bool:
        txt = _htmlish_to_text(txt_in)
        low = str(txt or "").lower()
        if require_convert and "convert" not in low:
            return False
        score = 0
        maturity_txt = str(row_in.get("maturity_display") or row_in.get("maturity_year") or "").strip().lower()
        if maturity_txt and maturity_txt in low:
            score += 3
        year_match = re.search(r"(20\d{2})", maturity_txt)
        if year_match and year_match.group(1) in low:
            score += 2
        coupon_num = pd.to_numeric(row_in.get("coupon_pct"), errors="coerce")
        coupon_tokens: set[str] = set()
        if pd.notna(coupon_num):
            coupon_val = float(coupon_num)
            if abs(coupon_val) <= 1.0:
                coupon_val *= 100.0
            coupon_tokens.update(
                {
                    f"{coupon_val:.2f}%".rstrip("0").rstrip("."),
                    f"{coupon_val:.2f}%",
                    f"{coupon_val:.1f}%".rstrip("0").rstrip("."),
                }
            )
        if any(tok and tok.lower() in low for tok in coupon_tokens):
            score += 3
        principal_val = pd.to_numeric(row_in.get("amount_principal"), errors="coerce")
        if pd.notna(principal_val):
            principal_m = float(principal_val)
            if abs(principal_m) > 1e6:
                principal_m /= 1e6
            principal_tokens = {
                f"${principal_m:.1f} million",
                f"{principal_m:.1f} million",
                f"${principal_m:.0f} million",
                f"{principal_m:.0f} million",
            }
            if any(tok.lower() in low for tok in principal_tokens):
                score += 2
        tranche_low = re.sub(r"[^a-z0-9% ]+", " ", str(row_in.get("tranche_name") or "").lower()).strip()
        if tranche_low and tranche_low in re.sub(r"[^a-z0-9% ]+", " ", low):
            score += 2
        return score >= 4
    def _repurchase_shares_from_source_docs(src_txt: Any, row_in: pd.Series) -> Optional[float]:
        single_convertible_context = bool(convert_df is not None and len(convert_df.index) == 1)
        src_txt_clean = ""
        try:
            if src_txt is not None and not pd.isna(src_txt):
                src_txt_clean = str(src_txt)
        except Exception:
            src_txt_clean = str(src_txt or "")
        for raw_path in src_txt_clean.split(" | "):
            doc_path = Path(str(raw_path).strip())
            if not doc_path.exists() or not doc_path.is_file():
                continue
            try:
                raw_text = doc_path.read_text(encoding="utf-8", errors="ignore")
            except Exception:
                continue
            txt = _htmlish_to_text(raw_text)
            rep_matches = list(
                re.finditer(
                    r"used approximately \$\s*([0-9][0-9,]*(?:\.\d+)?)\s*(million|billion|m|bn)?[^.]{0,240}?(?:to\s+)?repurchase(?: approximately)?\s*([0-9][0-9,]*(?:\.\d+)?)\s*(million|billion|m|bn)?\s+shares",
                    txt,
                    re.I,
                )
            )
            for rep_match in rep_matches:
                window = txt[max(0, rep_match.start() - 260): min(len(txt), rep_match.end() + 260)]
                if convert_df is not None and not convert_df.empty:
                    window_match_count = 0
                    for _, other_row in convert_df.iterrows():
                        if _convertible_text_matches_row(window, other_row, require_convert=False):
                            window_match_count += 1
                    if window_match_count > 1:
                        continue
                if not _convertible_text_matches_row(window, row_in, require_convert=False):
                    if not (single_convertible_context and _convertible_text_matches_row(txt, row_in, require_convert=True)):
                        continue
                try:
                    rep_shares = float(str(rep_match.group(3)).replace(",", ""))
                    rep_shares_unit = str(rep_match.group(4) or "").lower()
                    if rep_shares_unit in {"million", "m"}:
                        rep_shares *= 1e6
                    elif rep_shares_unit in {"billion", "bn"}:
                        rep_shares *= 1e9
                    return rep_shares
                except Exception:
                    continue
        return None
    has_convertible_rows = bool(convert_df is not None and not convert_df.empty)
    if not has_convertible_rows:
        no_conv_cell = ws.cell(row=convert_row, column=12, value="No convertible notes identified in latest debt set.")
        no_conv_cell.font = Font(size=11, italic=True, color="4B5563")
        no_conv_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        try:
            ws.merge_cells(start_row=convert_row, start_column=12, end_row=convert_row, end_column=convert_end_col)
        except Exception:
            pass
        for cc in range(12, convert_end_col + 1):
            ws.cell(row=convert_row, column=cc).border = thin_border
        ws.row_dimensions[convert_row].height = max(float(ws.row_dimensions[convert_row].height or 0), 18.0)
        convert_row += 1
    else:
        convert_total_rows: List[int] = []
        for _, conv_row in convert_df.iterrows():
            principal_m = pd.to_numeric(conv_row.get("amount_principal"), errors="coerce")
            coupon_val = pd.to_numeric(conv_row.get("coupon_pct"), errors="coerce")
            conv_price_val = pd.to_numeric(conv_row.get("conversion_price"), errors="coerce")
            conv_shares_val = pd.to_numeric(conv_row.get("shares_on_full_conversion"), errors="coerce")
            if pd.notna(conv_shares_val):
                conv_shares_val = float(conv_shares_val) / 1e6
            maturity_raw = conv_row.get("maturity_display") or conv_row.get("maturity_year") or ""
            maturity_txt = str(maturity_raw).strip()
            try:
                maturity_num = float(maturity_raw)
                if abs(maturity_num - round(maturity_num)) < 1e-9:
                    maturity_txt = str(int(round(maturity_num)))
            except Exception:
                pass
            coupon_pct = None
            if pd.notna(coupon_val):
                coupon_pct = float(coupon_val)
                if abs(coupon_pct) <= 1.0:
                    coupon_pct *= 100.0
            if coupon_pct is not None and maturity_txt:
                coupon_txt = f"{coupon_pct:.2f}".rstrip("0").rstrip(".")
                security_txt = f"{coupon_txt}% notes due {maturity_txt}"
            else:
                security_txt = re.sub(r"\bconvertible\b", "", str(conv_row.get("tranche_name") or ""), flags=re.I)
                security_txt = re.sub(r"\s+%", "%", security_txt)
                security_txt = re.sub(r"\s+", " ", security_txt).strip(" -")
            ws.cell(row=convert_row, column=12, value=security_txt)
            try:
                ws.merge_cells(start_row=convert_row, start_column=12, end_row=convert_row, end_column=13)
            except Exception:
                pass
            ws.cell(row=convert_row, column=14, value=(float(principal_m) / 1e6) if pd.notna(principal_m) else None).number_format = "#,##0.000"
            ws.cell(row=convert_row, column=15, value=float(coupon_val) / 100.0 if pd.notna(coupon_val) and float(coupon_val) > 1 else (float(coupon_val) if pd.notna(coupon_val) else None)).number_format = "0.00%"
            ws.cell(row=convert_row, column=16, value=conv_row.get("maturity_display") or conv_row.get("maturity_year"))
            ws.cell(row=convert_row, column=conv_price_col, value=float(conv_price_val) if pd.notna(conv_price_val) else None).number_format = "$#,##0.00"
            ws.cell(row=convert_row, column=shares_full_col, value=float(conv_shares_val) if pd.notna(conv_shares_val) else None).number_format = "#,##0.000"
            rep_shares_val = pd.to_numeric(conv_row.get("concurrent_repurchase_shares"), errors="coerce")
            if not pd.notna(rep_shares_val):
                rep_shares_val = pd.to_numeric(
                    _repurchase_shares_from_source_docs(conv_row.get("conversion_terms_source"), conv_row),
                    errors="coerce",
                )
            if pd.notna(rep_shares_val):
                rep_shares_val = float(rep_shares_val) / 1e6
            ws.cell(row=convert_row, column=concurrent_rep_col, value=float(rep_shares_val) if pd.notna(rep_shares_val) else None).number_format = "#,##0.000"
            ws.cell(
                row=convert_row,
                column=net_added_col,
                value=f'=IF({shares_full_letter}{convert_row}="","",{shares_full_letter}{convert_row}-IF({concurrent_rep_letter}{convert_row}="",0,{concurrent_rep_letter}{convert_row}))',
            ).number_format = "#,##0.000"
            dilution_note = _safe_text_value(conv_row.get("dilution_structure_note"))
            if dilution_note:
                comment_cell = ws.cell(row=convert_row, column=shares_full_col)
                if not pd.notna(conv_shares_val):
                    comment_cell = ws.cell(row=convert_row, column=12)
                _set_cell_comment_local(comment_cell, dilution_note)
            note_txt = _safe_text_value(conv_row.get("conversion_terms_note"))
            src_txt = _safe_text_value(conv_row.get("conversion_terms_source"))
            action_note_bits: List[str] = []
            rep_amt = pd.to_numeric(conv_row.get("concurrent_repurchase_amount"), errors="coerce")
            rep_sh = pd.to_numeric(conv_row.get("concurrent_repurchase_shares"), errors="coerce")
            if pd.notna(rep_amt):
                action_note_bits.append(f"Concurrent repurchase: ${float(rep_amt)/1e6:,.1f}m")
            if pd.notna(rep_sh):
                action_note_bits.append(f"Concurrent repurchase shares: {float(rep_sh)/1e6:,.1f}m")
            settlement_txt = _safe_text_value(conv_row.get("settlement_type"))
            hedge_txt = _safe_text_value(conv_row.get("hedge_or_call_spread"))
            conditions_txt = _safe_text_value(conv_row.get("conversion_conditions_note"))
            if settlement_txt:
                action_note_bits.append(f"Settlement: {settlement_txt}")
            if hedge_txt:
                action_note_bits.append(f"Hedge/call spread: {hedge_txt}")
            if conditions_txt:
                action_note_bits.append(f"Conditions: {conditions_txt}")
            if note_txt or src_txt or action_note_bits:
                try:
                    comment_txt = "\n".join([x for x in [note_txt] + action_note_bits if x])
                    if src_txt:
                        comment_txt = f"{comment_txt}\n\nSource: {src_txt}" if comment_txt else f"Source: {src_txt}"
                    _set_cell_comment_local(ws.cell(row=convert_row, column=12), comment_txt)
                except Exception:
                    pass
            for cc in range(12, convert_end_col + 1):
                ws.cell(row=convert_row, column=cc).border = thin_border
                ws.cell(row=convert_row, column=cc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            convert_total_rows.append(convert_row)
            convert_row += 1
        if len(convert_total_rows) > 1:
            ws.cell(row=convert_row, column=12, value="Total").font = bold
            for cc in (14, shares_full_col, concurrent_rep_col, net_added_col):
                col_letter = get_column_letter(cc)
                ws.cell(row=convert_row, column=cc, value=f"=SUM({col_letter}{convert_total_rows[0]}:{col_letter}{convert_total_rows[-1]})").font = bold
                ws.cell(row=convert_row, column=cc).number_format = "#,##0.000"
            for cc in range(12, convert_end_col + 1):
                ws.cell(row=convert_row, column=cc).border = thin_border
            convert_row += 1
    ws.cell(row=convert_row + 1, column=12, value="When valuing as-converted scenarios, adjust BOTH share count and net debt.")
    try:
        ws.merge_cells(start_row=convert_row + 1, start_column=12, end_row=convert_row + 1, end_column=convert_end_col)
    except Exception:
        pass
    ws.cell(row=convert_row + 1, column=12).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for cc in range(12, convert_end_col + 1):
        ws.cell(row=convert_row + 1, column=cc).border = thin_border
    ws.row_dimensions[convert_row + 1].height = 30
    ws.cell(
        row=convert_row + 2,
        column=12,
        value="Illustrative conversion cases often use 50% / 70% / 90% of net added shares; adjust both share count and net debt when modeling as-converted outcomes.",
    )
    try:
        ws.merge_cells(start_row=convert_row + 2, start_column=12, end_row=convert_row + 2, end_column=convert_end_col)
    except Exception:
        pass
    ws.cell(row=convert_row + 2, column=12).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for cc in range(12, convert_end_col + 1):
        ws.cell(row=convert_row + 2, column=cc).border = thin_border
    ws.row_dimensions[convert_row + 2].height = 30
    if not has_convertible_rows:
        # Keep the convertible-note section visible, but avoid the
        # as-converted modeling notes when no convertible instrument exists.
        for rr_clear in (convert_row + 1, convert_row + 2):
            for merged in list(ws.merged_cells.ranges):
                try:
                    min_col, min_row, max_col, max_row = merged.bounds
                except Exception:
                    continue
                if max_row < rr_clear or min_row > rr_clear or max_col < 12 or min_col > convert_end_col:
                    continue
                try:
                    ws.unmerge_cells(str(merged))
                except Exception:
                    pass
            for cc_clear in range(12, convert_end_col + 1):
                c_clear = ws.cell(rr_clear, cc_clear)
                c_clear.value = None
                c_clear.comment = None
                c_clear.fill = PatternFill(fill_type=None)
                c_clear.border = Border()
                c_clear.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            ws.row_dimensions[rr_clear].height = max(float(ws.row_dimensions[rr_clear].height or 0), 18.0)

    # Compact optional DCF module (display-only add-on; core valuation formulas unchanged).
    ws.cell(row=row_dcf_hdr, column=dcf_label_col, value="DCF (optional module)").font = bold
    ws.cell(row=row_dcf_hdr, column=dcf_label_col).fill = section_fill
    ws.cell(row=row_dcf_hdr, column=dcf_interp_col, value="Interpretation").font = bold
    ws.cell(row=row_dcf_hdr, column=dcf_interp_col).fill = header_fill
    _box(
        row_dcf_start,
        "Starting FCFF ($m) [ASSUMPTION]",
        "=IF(Adj_FCF_TTM<>\"\",Adj_FCF_TTM,IF(FCF_TTM<>\"\",FCF_TTM,\"\"))",
        True,
        "#,##0.000",
        dcf_label_col,
        dcf_value_col,
    )
    _box(row_dcf_g, "g (Years 1-5) [ASSUMPTION]", 0.01, True, "0.0%", dcf_label_col, dcf_value_col)
    _box(row_dcf_gt, "Terminal g [ASSUMPTION]", 0.03, True, "0.0%", dcf_label_col, dcf_value_col)
    _box(row_dcf_wacc, "WACC [ASSUMPTION]", 0.10, True, "0.0%", dcf_label_col, dcf_value_col)
    _box(
        row_dcf_ev,
        "DCF EV ($m)",
        (
            f"=IF(OR({dcf_col}{row_dcf_start}=\"\",{dcf_col}{row_dcf_start}<=0,{dcf_col}{row_dcf_wacc}<=0,{dcf_col}{row_dcf_wacc}<={dcf_col}{row_dcf_gt}),\"\"," 
            f"({dcf_col}{row_dcf_start}*(1+{dcf_col}{row_dcf_g})/({dcf_col}{row_dcf_wacc}-{dcf_col}{row_dcf_g})*(1-((1+{dcf_col}{row_dcf_g})/(1+{dcf_col}{row_dcf_wacc}))^5)+"
            f"({dcf_col}{row_dcf_start}*(1+{dcf_col}{row_dcf_g})^5*(1+{dcf_col}{row_dcf_gt})/({dcf_col}{row_dcf_wacc}-{dcf_col}{row_dcf_gt})/((1+{dcf_col}{row_dcf_wacc})^5)))"
            ")"
        ),
        False,
        "#,##0.000",
        dcf_label_col,
        dcf_value_col,
    )
    _box(
        row_dcf_eq,
        "DCF Equity value/share",
        f"=IF(OR({dcf_col}{row_dcf_ev}=\"\",{fair_denom}=\"\",{fair_denom}<=0),\"\",({dcf_col}{row_dcf_ev}-NetDebt)/{fair_denom})",
        False,
        "$#,##0.00",
        dcf_label_col,
        dcf_value_col,
    )
    _set_interpretation(row_dcf_start, "Default = Adj FCF TTM, fallback to FCF TTM.", dcf_interp_col)
    _set_interpretation(row_dcf_g, "Short-horizon growth assumption for years 1-5.", dcf_interp_col)
    _set_interpretation(row_dcf_gt, "Perpetual growth assumption used in terminal value.", dcf_interp_col)
    _set_interpretation(row_dcf_wacc, "Discount rate assumption; higher WACC lowers DCF value.", dcf_interp_col)
    _set_interpretation(row_dcf_ev, "Enterprise value from 5-year growth + terminal value.", dcf_interp_col)
    _set_interpretation(row_dcf_eq, "DCF EV less net debt, divided by selected share denominator.", dcf_interp_col)
    _set_cell_comment(ws.cell(row=row_dcf_start, column=dcf_value_col), "Assumption input.")
    _set_cell_comment(ws.cell(row=row_dcf_g, column=dcf_value_col), "Assumption input.")
    _set_cell_comment(ws.cell(row=row_dcf_gt, column=dcf_value_col), "Assumption input.")
    _set_cell_comment(ws.cell(row=row_dcf_wacc, column=dcf_value_col), "Assumption input.")

    # DCF sensitivity on WACC vs terminal g.
    dcf_sens_col0 = dcf_label_col
    dcf_sens_value_cols = [dcf_sens_col0 + 1 + _i for _i in range(len(dcf_sens_g_vals))]
    ws.cell(row=row_dcf_sens_hdr, column=dcf_sens_col0, value="DCF Sensitivity ($/share)").font = bold
    ws.cell(row=row_dcf_sens_hdr, column=dcf_sens_col0).fill = section_fill
    try:
        ws.merge_cells(
            start_row=row_dcf_sens_hdr,
            start_column=dcf_sens_col0,
            end_row=row_dcf_sens_hdr,
            end_column=dcf_sens_col0 + len(dcf_sens_g_vals),
        )  # G:L
    except Exception:
        pass
    ws.cell(row=row_dcf_sens_hdr + 1, column=dcf_sens_col0, value="WACC \\ gT").font = bold
    for _i, _g in enumerate(dcf_sens_g_vals):
        c = ws.cell(row=row_dcf_sens_hdr + 1, column=dcf_sens_value_cols[_i], value=_g)
        c.font = bold
        c.number_format = "0.0%"
    for _j, _w in enumerate(dcf_sens_wacc_vals):
        rr = row_dcf_sens_hdr + 2 + _j
        c_w = ws.cell(row=rr, column=dcf_sens_col0, value=_w)
        c_w.font = bold
        c_w.number_format = "0.0%"
        for _i, _c in enumerate(dcf_sens_value_cols):
            c_gt = ws.cell(row=row_dcf_sens_hdr + 1, column=_c)
            w_ref = f"${get_column_letter(dcf_sens_col0)}${rr}"
            gt_ref = f"${get_column_letter(c_gt.column)}${c_gt.row}"
            ws.cell(
                row=rr,
                column=_c,
                value=(
                    f"=IF(OR({dcf_col}{row_dcf_start}=\"\",{dcf_col}{row_dcf_start}<=0,{w_ref}<=0,{w_ref}<={gt_ref},{fair_denom}=\"\",{fair_denom}<=0),\"\"," 
                    f"(({dcf_col}{row_dcf_start}*(1+{dcf_col}{row_dcf_g})/({w_ref}-{dcf_col}{row_dcf_g})*(1-((1+{dcf_col}{row_dcf_g})/(1+{w_ref}))^5)+"
                    f"({dcf_col}{row_dcf_start}*(1+{dcf_col}{row_dcf_g})^5*(1+{gt_ref})/({w_ref}-{gt_ref})/((1+{w_ref})^5))-NetDebt)/{fair_denom}))"
                ),
            ).number_format = "$#,##0.00"
    # Keep DCF sensitivity cell alignment consistent (no wrap / top-left drift).
    for cc in dcf_sens_value_cols:
        ws.cell(row=row_dcf_sens_hdr + 1, column=cc).alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
    for rr in range(row_dcf_sens_hdr + 2, row_dcf_sens_last_row + 1):
        ws.cell(row=rr, column=dcf_sens_col0).alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
        for cc in dcf_sens_value_cols:
            ws.cell(row=rr, column=cc).alignment = Alignment(horizontal="right", vertical="bottom", wrap_text=False)
    # Keep styles aligned across the expanded sensitivity matrix.
    for rr in range(row_dcf_sens_hdr + 1, row_dcf_sens_last_row + 1):
        for cc in dcf_sens_value_cols[1:]:
            try:
                ws.cell(row=rr, column=cc)._style = copy(ws.cell(row=rr, column=dcf_sens_value_cols[0])._style)
            except Exception:
                pass
    # Re-apply alignment after style-copy so all gT columns keep the same layout.
    for cc in dcf_sens_value_cols:
        ws.cell(row=row_dcf_sens_hdr + 1, column=cc).alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
    for rr in range(row_dcf_sens_hdr + 2, row_dcf_sens_last_row + 1):
        for cc in dcf_sens_value_cols:
            ws.cell(row=rr, column=cc).alignment = Alignment(horizontal="right", vertical="bottom", wrap_text=False)
    # Highlight "reasonable zone": WACC 9-13% and gT 1-3% -> I227:K231.
    try:
        zone_fill = PatternFill("solid", fgColor="D9EAD3")
        thin_side = Side(style="thin", color="000000")
        med_side = Side(style="medium", color="000000")
        zone_r1, zone_r2 = row_dcf_sens_hdr + 3, row_dcf_sens_hdr + 7  # 227:231
        zone_c1, zone_c2 = dcf_sens_col0 + 2, dcf_sens_col0 + 4        # I:K
        # Bonus highlight headers/labels for readability.
        for cc in range(zone_c1, zone_c2 + 1):
            ws.cell(row=row_dcf_sens_hdr + 1, column=cc).fill = zone_fill
        for rr in range(zone_r1, zone_r2 + 1):
            ws.cell(row=rr, column=dcf_sens_col0).fill = zone_fill
        for rr in range(zone_r1, zone_r2 + 1):
            for cc in range(zone_c1, zone_c2 + 1):
                c = ws.cell(row=rr, column=cc)
                c.fill = zone_fill
                c.border = Border(
                    left=med_side if cc == zone_c1 else thin_side,
                    right=med_side if cc == zone_c2 else thin_side,
                    top=med_side if rr == zone_r1 else thin_side,
                    bottom=med_side if rr == zone_r2 else thin_side,
                )
    except Exception:
        pass

    # Explanatory helper box beside market-implied block.
    try:
        ws.merge_cells(start_row=222, start_column=20, end_row=231, end_column=26)  # T:Z
    except Exception:
        pass
    expl_cell = ws.cell(
        row=222,
        column=20,
        value=(
            "Högre WACC = marknaden kräver mer betalt för risken => framtida cash flow värderas lägre idag.\n\n"
            "gT (terminaltillväxt): antagen långsiktig FCF tillväxt (efter prognosperioden).\n\n"
            "Market-implied gT (givet WACC) = vad priset kräver att du tror.\n"
            "gT < 0% => marknaden prisar krympning/erosion\n"
            "gT ~1-3% => normal långsiktig nominell tillväxt\n"
            "gT >3-4% => marknaden kräver stark uthållig growth (eller WACC/FCF-antaganden är för snälla)."
        ),
    )
    expl_cell.fill = copy(section_fill)
    expl_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    expl_cell.font = Font(size=max(9, font_size - 1))
    for rr in range(222, 232):
        ws.row_dimensions[rr].height = max(ws.row_dimensions[rr].height or 0, 22)

    # Re-assert labels/headers after all formula writes to keep the valuation panel readable.
    ws.cell(row=valuation_header_row, column=input_label_col, value="Valuation")
    ws.cell(row=valuation_header_row, column=input_label_col).font = bold
    ws.cell(row=valuation_header_row, column=input_label_col).fill = section_fill
    ws.cell(row=valuation_inputs_row, column=input_label_col, value="Inputs").font = bold
    ws.cell(row=valuation_inputs_row, column=input_basis_col, value="Basis").font = bold
    ws.cell(row=valuation_inputs_row, column=input_hint_col, value="Hint").font = bold
    ws.cell(row=valuation_inputs_row, column=output_label_col, value="Outputs").font = bold
    ws.cell(row=valuation_inputs_row, column=output_interp_col, value="Interpretation").font = bold
    for cc in [input_label_col, input_basis_col, input_hint_col, output_label_col, output_interp_col]:
        ws.cell(row=valuation_inputs_row, column=cc).fill = header_fill
    ws.cell(row=row_market_hdr, column=market_label_col, value="What Market Is Pricing").font = bold
    ws.cell(row=row_market_hdr, column=market_label_col).fill = section_fill

    for rr, lbl in [
        (row_price, "Price"),
        (row_asof, "As of"),
        (row_shares_out, "Shares outstanding (m)"),
        (row_shares_dil, "Shares diluted (m)"),
        (row_net_debt, "Net debt (core, $m)"),
        (row_ebitda_ttm, "EBITDA TTM ($m)"),
        (row_adj_ebitda_ttm, "Adj EBITDA TTM ($m)"),
        (row_fcf_ttm, "FCF TTM ($m)"),
        (row_adj_fcf_ttm, "Adj FCF TTM ($m)"),
        (row_rev_ttm, "Revenue TTM ($m)"),
        (row_eps_ttm, "EPS TTM ($)"),
        (row_adj_eps_ttm, "Adj EPS TTM ($)"),
        (row_bv, "BV/share"),
        (row_tbv, "TBV/share"),
        (row_tgt_ev_adj, "Target EV/Adj EBITDA"),
        (row_tgt_ev, "Target EV/EBITDA"),
        (row_tgt_fcf, "Target EV yield"),
        (row_capex_ttm, "Capex TTM ($m)"),
        (row_int_paid_ttm, "Interest paid TTM ($m)"),
        (row_owner_maint_ratio, "Maint. capex % of capex"),
        (row_owner_recurring, "Recurring cash costs ($m)"),
        (row_owner_wc_norm, "WC normalization ($m)"),
        (row_share_mode, "Per-share denominator"),
    ]:
        ws.cell(row=rr, column=input_label_col, value=lbl)

    for rr, lbl in [
        (row_mktcap, "Market cap ($m)"),
        (row_ev, "EV ($m)"),
        (row_implied_ev_adj, "Implied EV/Adj EBITDA"),
        (row_implied_ev, "Implied EV/EBITDA"),
        (row_fcff_proxy_ttm, "FCFF proxy TTM ($m)"),
        (row_implied_fcff, "Implied FCFF yield (EV)"),
        (row_equity_fcf, "Equity FCF yield"),
        (row_owner_fcf_ttm, "Owner earnings TTM ($m)"),
        (row_owner_fcf_yield, "Owner earnings yield"),
        (row_eq_adj, "Eq/Share @ target EV/Adj EBITDA"),
        (row_eq_ev, "Eq/Share @ target EV/EBITDA"),
        (row_eq_fcf, "Eq/Share @ target EV yield (FCFF)"),
        (row_pe, "P/E (TTM)"),
        (row_pe_adj, "P/E (Adj TTM)"),
        (row_ev_sales, "EV/Sales (TTM)"),
        (row_pb, "Price/BV"),
        (row_ptbv, "Price/TBV"),
    ]:
        ws.cell(row=rr, column=output_label_col, value=lbl)

    ws.cell(row=row_market_hdr, column=market_label_col, value="What Market Is Pricing")
    for rr, lbl in [
        (row_req_adj_ebitda, "Required Adj EBITDA @ target multiple ($m)"),
        (row_req_adj_delta, "Implied Adj EBITDA change vs current"),
        (row_req_fcff, "Required FCFF @ target EV yield ($m)"),
        (row_req_fcff_delta, "Implied FCFF change vs current"),
        (row_req_owner_fcf, "Required owner earnings @ target EV yield ($m)"),
        (row_req_owner_delta, "Implied owner earnings change"),
    ]:
        ws.cell(row=rr, column=market_label_col, value=lbl)
    for rr, lbl in [
        (row_scn_hdr, "Trigger Scenarios"),
        (row_scn_profile, "Scenario preset (Base/Bull/Bear/Custom)"),
        (row_scn_growth, "Rev growth assumption"),
        (row_scn_margin, "Adj EBITDA margin assumption"),
        (row_scn_refi, "Refi interest normalization ($m)"),
        (row_scn_buyback, "Buyback continuation (m shares)"),
        (row_scn_adj_ebitda, "Scenario Adj EBITDA ($m)"),
        (row_scn_owner_fcf, "Scenario owner earnings ($m)"),
        (row_scn_eq_ev, "Eq/share scenario (EV/Adj)"),
        (row_scn_eq_fcf, "Eq/share scenario (EV owner earnings yield)"),
    ]:
        ws.cell(row=rr, column=scn_label_col, value=lbl)
    for rr, lbl in [
        (row_drv_hdr, "Scenario drivers (internal)"),
        (row_drv_rev, "Revenue trend driver"),
        (row_drv_margin, "Margin driver"),
        (row_drv_fcf, "FCF conversion driver"),
        (row_drv_lev, "Leverage/cap allocation driver"),
    ]:
        ws.cell(row=rr, column=driver_label_col, value=lbl)
    for rr, lbl in [
        (row_toggle_hdr, "Quality toggles"),
        (row_toggle_reg, "Regulatory dependency?"),
        (row_toggle_gap, "Persistent non-GAAP gap?"),
        (row_toggle_lev, "Balance-sheet optionality constrained?"),
        (row_toggle_conc, "Customer concentration disclosed?"),
    ]:
        ws.cell(row=rr, column=toggle_label_col, value=lbl)
    for rr, lbl in [
        (row_qadj_hdr, "Quality-adjusted target multiple/yield"),
        (row_qadj_ev_adj, "Adj target EV/Adj EBITDA (x)"),
        (row_qadj_ev, "Adj target EV/EBITDA (x)"),
        (row_qadj_yield, "Adj target EV yield"),
    ]:
        ws.cell(row=rr, column=qadj_label_col, value=lbl)
    for rr, lbl in [
        (row_dcf_hdr, "DCF (optional module)"),
        (row_dcf_start, "Starting FCFF ($m) [ASSUMPTION]"),
        (row_dcf_g, "g (Years 1-5) [ASSUMPTION]"),
        (row_dcf_gt, "Terminal g [ASSUMPTION]"),
        (row_dcf_wacc, "WACC [ASSUMPTION]"),
        (row_dcf_ev, "DCF EV ($m)"),
        (row_dcf_eq, "DCF Equity value/share"),
    ]:
        ws.cell(row=rr, column=dcf_label_col, value=lbl)

    # QA check (latest quarter)
    if date_ref is not None:
        try:
            if (fcf_ttm is None) or (int_paid_ttm_val is None):
                info_log = pd.concat(
                    [
                        info_log,
                        pd.DataFrame(
                            [
                                {
                                    "quarter": pd.Timestamp(date_ref).date(),
                                    "metric": "valuation_fcff_proxy_guard",
                                    "severity": "warn",
                                    "message": "FCFF_Proxy_TTM formula guard active: missing/invalid FCF_TTM or InterestPaid_TTM.",
                                    "source": "Valuation",
                                }
                            ]
                        ),
                    ],
                    ignore_index=True,
                )
        except Exception:
            pass
    qa_msgs = []
    if date_ref is not None:
        q = date_ref
        def _chk(label: str, val_m: Optional[float], src: Optional[float]) -> None:
            if val_m is None or src is None:
                return
            if abs(val_m - (src / 1e6)) > 0.001:
                qa_msgs.append(label)
        _chk("Revenue", rev_map.get(q), rev_map.get(q))
        _chk("EBITDA", ebitda_map.get(q), ebitda_map.get(q))
        _chk("CFO", cfo_map.get(q), cfo_map.get(q))
        _chk("Capex", capex_map.get(q), capex_map.get(q))
        _chk("Cash", cash_map.get(q), cash_map.get(q))
        _chk("Debt core", debt_core_map.get(q), debt_core_map.get(q))
    if date_ref is not None:
        if rev_map.get(date_ref) is None:
            qa_msgs.append("Revenue missing")
        if ebitda_map.get(date_ref) is None:
            qa_msgs.append("EBITDA missing")
        if cfo_map.get(date_ref) is None:
            qa_msgs.append("CFO missing")
        if capex_map.get(date_ref) is None:
            qa_msgs.append("Capex missing")
        if cash_map.get(date_ref) is None:
            qa_msgs.append("Cash missing")
        if debt_core_map.get(date_ref) is None:
            qa_msgs.append("Debt core missing")
    if tieout_diff_m is not None and abs(tieout_diff_m) > 50.0:
        qa_msgs.append("Debt principal/carrying recon > $50m")
    if qa_msgs:
        ws["A4"] = "QA: " + ", ".join(qa_msgs)
        ws["A4"].font = bold


    return ValuationFormulaCoreRenderResult(
        valuation_header_row=valuation_header_row,
        valuation_inputs_row=valuation_inputs_row,
        input_label_col=input_label_col,
        input_value_col=input_value_col,
        input_basis_col=input_basis_col,
        input_hint_col=input_hint_col,
        output_label_col=output_label_col,
        output_value_col=output_value_col,
        output_interp_col=output_interp_col,
        market_label_col=market_label_col,
        market_value_col=market_value_col,
        market_interp_col=market_interp_col,
        scn_label_col=scn_label_col,
        scn_value_col=scn_value_col,
        scn_interp_col=scn_interp_col,
        driver_label_col=driver_label_col,
        driver_value_col=driver_value_col,
        toggle_label_col=toggle_label_col,
        toggle_value_col=toggle_value_col,
        qadj_label_col=qadj_label_col,
        qadj_value_col=qadj_value_col,
        qadj_text_col=qadj_text_col,
        dcf_label_col=dcf_label_col,
        dcf_value_col=dcf_value_col,
        dcf_interp_col=dcf_interp_col,
        grid_start=grid_start,
        grid_layout_width=grid_layout_width,
        right_stack_anchor=right_stack_anchor,
        date_ref=date_ref,
        row_price=row_price,
        row_asof=row_asof,
        row_shares_out=row_shares_out,
        row_shares_dil=row_shares_dil,
        row_net_debt=row_net_debt,
        row_ebitda_ttm=row_ebitda_ttm,
        row_adj_ebitda_ttm=row_adj_ebitda_ttm,
        row_fcf_ttm=row_fcf_ttm,
        row_adj_fcf_ttm=row_adj_fcf_ttm,
        row_rev_ttm=row_rev_ttm,
        row_eps_ttm=row_eps_ttm,
        row_adj_eps_ttm=row_adj_eps_ttm,
        row_bv=row_bv,
        row_tbv=row_tbv,
        row_tgt_ev_adj=row_tgt_ev_adj,
        row_tgt_ev=row_tgt_ev,
        row_tgt_fcf=row_tgt_fcf,
        row_capex_ttm=row_capex_ttm,
        row_int_paid_ttm=row_int_paid_ttm,
        row_owner_maint_ratio=row_owner_maint_ratio,
        row_owner_recurring=row_owner_recurring,
        row_owner_wc_norm=row_owner_wc_norm,
        row_share_mode=row_share_mode,
        row_out_hdr=row_out_hdr,
        row_mktcap=row_mktcap,
        row_ev=row_ev,
        row_implied_ev_adj=row_implied_ev_adj,
        row_implied_ev=row_implied_ev,
        row_fcff_proxy_ttm=row_fcff_proxy_ttm,
        row_implied_fcff=row_implied_fcff,
        row_equity_fcf=row_equity_fcf,
        row_owner_fcf_ttm=row_owner_fcf_ttm,
        row_owner_fcf_yield=row_owner_fcf_yield,
        row_eq_adj=row_eq_adj,
        row_eq_ev=row_eq_ev,
        row_eq_fcf=row_eq_fcf,
        row_pe=row_pe,
        row_pe_adj=row_pe_adj,
        row_ev_sales=row_ev_sales,
        row_pb=row_pb,
        row_ptbv=row_ptbv,
        row_mi_hdr=row_mi_hdr,
        row_mi_market_ev=row_mi_market_ev,
        row_mi_dcf_ev=row_mi_dcf_ev,
        row_mi_curr_wacc=row_mi_curr_wacc,
        row_mi_curr_gt=row_mi_curr_gt,
        row_mi_tbl_hdr=row_mi_tbl_hdr,
        row_mi_wacc_start=row_mi_wacc_start,
        row_mi_wacc_end=row_mi_wacc_end,
        row_mi_toggle=row_mi_toggle,
        row_dcf_hdr=row_dcf_hdr,
        row_dcf_start=row_dcf_start,
        row_dcf_g=row_dcf_g,
        row_dcf_gt=row_dcf_gt,
        row_dcf_wacc=row_dcf_wacc,
        row_dcf_ev=row_dcf_ev,
        row_dcf_eq=row_dcf_eq,
        row_dcf_sens_hdr=row_dcf_sens_hdr,
        row_dcf_sens_last_row=row_dcf_sens_last_row,
        row_scn_hdr=row_scn_hdr,
        row_scn_profile=row_scn_profile,
        row_scn_growth=row_scn_growth,
        row_scn_margin=row_scn_margin,
        row_scn_refi=row_scn_refi,
        row_scn_buyback=row_scn_buyback,
        row_scn_adj_ebitda=row_scn_adj_ebitda,
        row_scn_owner_fcf=row_scn_owner_fcf,
        row_scn_eq_ev=row_scn_eq_ev,
        row_scn_eq_fcf=row_scn_eq_fcf,
        row_market_hdr=row_market_hdr,
        row_req_adj_ebitda=row_req_adj_ebitda,
        row_req_adj_delta=row_req_adj_delta,
        row_req_fcff=row_req_fcff,
        row_req_fcff_delta=row_req_fcff_delta,
        row_req_owner_fcf=row_req_owner_fcf,
        row_req_owner_delta=row_req_owner_delta,
        row_qa=row_qa,
        row_drv_hdr=row_drv_hdr,
        row_drv_rev=row_drv_rev,
        row_drv_margin=row_drv_margin,
        row_drv_fcf=row_drv_fcf,
        row_drv_lev=row_drv_lev,
        row_toggle_hdr=row_toggle_hdr,
        row_toggle_reg=row_toggle_reg,
        row_toggle_gap=row_toggle_gap,
        row_toggle_lev=row_toggle_lev,
        row_toggle_conc=row_toggle_conc,
        row_qadj_hdr=row_qadj_hdr,
        row_qadj_ev_adj=row_qadj_ev_adj,
        row_qadj_ev=row_qadj_ev,
        row_qadj_yield=row_qadj_yield,
        row_dcf_end=row_dcf_end,
        row_hv_hdr=row_hv_hdr,
        row_hv_total=row_hv_total,
        row_hv_prof=row_hv_prof,
        row_hv_cash=row_hv_cash,
        row_hv_delev=row_hv_delev,
        row_hv_quality=row_hv_quality,
        row_hv_narr=row_hv_narr,
        row_hv_b1=row_hv_b1,
        row_hv_b2=row_hv_b2,
        row_hv_b3=row_hv_b3,
        row_hv_b4=row_hv_b4,
        row_hv_b5=row_hv_b5,
        row_convert_hdr=row_convert_hdr,
        convert_header_end_col=convert_header_end_col,
        qa_msgs=qa_msgs,
        tieout_diff_m=tieout_diff_m,
        fair_denom=fair_denom,
        normalize_thesis_bridge_basis=_normalize_thesis_bridge_basis,
        set_formula_name=_set_formula_name,
        set_cell_comment=_set_cell_comment,
    )
