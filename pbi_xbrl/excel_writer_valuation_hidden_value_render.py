"""Worksheet render adapter for the Valuation Hidden Value visible panel."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Mapping, MutableMapping, Optional

import pandas as pd
from openpyxl.styles import Alignment, Border, PatternFill
from openpyxl.utils import get_column_letter

from .excel_writer_hidden_value_surface import (
    HiddenValueSurfaceModelInputs,
    NO_TRIGGER_DISPLAY_LABEL,
    NO_TRIGGER_DISPLAY_SCORE,
    NO_TRIGGER_DISPLAY_SEVERITY,
    NO_TRIGGER_DISPLAY_SUPPORT,
    NO_TRIGGER_DISPLAY_TITLE,
    build_hidden_value_surface_model,
    hidden_flag_field,
    hidden_flag_score,
    hidden_value_ai_helper_formula,
)


@dataclass(frozen=True)
class ValuationHiddenValueRenderDeps:
    runtime: MutableMapping[str, Any]


@dataclass(frozen=True)
class ValuationHiddenValueRenderResult:
    row_hidden_value_start: int
    row_hidden_value_end: int
    row_flags_header: int
    row_flags_columns: int
    row_flags_start: int
    row_flags_end: int
    row_operating_signals_header: int
    row_capital_return_header: int
    row_buybacks: int
    row_buybacks_note: int
    row_dividends: int
    row_dividends_note: int
    row_score_panel_header: int
    next_panel_row: int
    helper_column: int
    visible_flag_count: int
    valuation_export_expectation: Mapping[str, Any] | None


def render_valuation_hidden_value_panel(
    deps: ValuationHiddenValueRenderDeps,
) -> ValuationHiddenValueRenderResult:
    __rt = deps.runtime
    context_globals = dict(__rt.get("context_globals") or {})

    def _rt_get(name: str) -> Any:
        if name in __rt:
            return __rt[name]
        if name in context_globals:
            return context_globals[name]
        return globals().get(name)

    _anf_visible_quarter_label = _rt_get("_anf_visible_quarter_label")
    _build_hidden_value_flags_fallback = _rt_get("_build_hidden_value_flags_fallback")
    _estimate_wrapped_row_height = _rt_get("_estimate_wrapped_row_height")
    _fmt_short_money_value_local = _rt_get("_fmt_short_money_value_local")
    adj_ebit_ttm_map = _rt_get("adj_ebit_ttm_map")
    adj_ebitda_ttm_map = _rt_get("adj_ebitda_ttm_map")
    adj_metrics = _rt_get("adj_metrics")
    bold = _rt_get("bold")
    build_hidden_value_flags = _rt_get("build_hidden_value_flags")
    buyback_map = _rt_get("buyback_map")
    buyback_ttm_map = _rt_get("buyback_ttm_map")
    cov_cash_display_map = _rt_get("cov_cash_display_map")
    cov_pnl_display_map = _rt_get("cov_pnl_display_map")
    cov_pnl_map = _rt_get("cov_pnl_map")
    ctx_ref = _rt_get("ctx_ref")
    debt_tranches = _rt_get("debt_tranches")
    dividend_ttm_map = _rt_get("dividend_ttm_map")
    flags_audit_df = _rt_get("flags_audit_df")
    flags_df = _rt_get("flags_df")
    glx_normalize_text = _rt_get("glx_normalize_text")
    header_fill = _rt_get("header_fill")
    hist = _rt_get("hist")
    hv_buybacks = _rt_get("hv_buybacks")
    hv_buybacks_note = _rt_get("hv_buybacks_note")
    hv_dividends = _rt_get("hv_dividends")
    hv_dividends_note = _rt_get("hv_dividends_note")
    hv_obs = _rt_get("hv_obs")
    hv_scores = _rt_get("hv_scores")
    is_anf_profile = _rt_get("is_anf_profile")
    leverage_df = _rt_get("leverage_df")
    net_lev_adj_display_map = _rt_get("net_lev_adj_display_map")
    net_lev_adj_map = _rt_get("net_lev_adj_map")
    net_lev_display_map = _rt_get("net_lev_display_map")
    price = _rt_get("price")
    qs_ts = _rt_get("qs_ts")
    section_fill = _rt_get("section_fill")
    signals_base_df = _rt_get("signals_base_df")
    thin_border = _rt_get("thin_border")
    valuation_header_row = _rt_get("valuation_header_row")
    ws = _rt_get("ws")

    # Left-side visible flags / operating signals / capital return in columns A:K.
    hv_label_col = 1  # A
    hv_val_col = 2    # B
    hv_panel_label_col = 14  # N
    hv_panel_label_end_col = 17  # Q
    hv_panel_val_col = 18  # R
    hv_panel_body_fill = PatternFill("solid", fgColor="F7FAFC")

    def _box_hv(row_idx: int, label: str, value: Any = None, number_format: Optional[str] = None) -> None:
        label_cell = ws.cell(row=row_idx, column=hv_label_col, value=label)
        label_cell.font = bold
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        vcell = ws.cell(row=row_idx, column=hv_val_col, value=value)
        if number_format:
            vcell.number_format = number_format

    def _box_hv_panel(row_idx: int, label: str, value: Any = None, number_format: Optional[str] = None) -> None:
        try:
            ws.merge_cells(
                start_row=row_idx,
                start_column=hv_panel_label_col,
                end_row=row_idx,
                end_column=hv_panel_label_end_col,
            )
        except Exception:
            pass
        lbl = ws.cell(row=row_idx, column=hv_panel_label_col, value=label)
        lbl.font = bold
        lbl.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        vcell = ws.cell(row=row_idx, column=hv_panel_val_col, value=value)
        if number_format:
            vcell.number_format = number_format
        vcell.alignment = Alignment(
            horizontal="right" if number_format else "left",
            vertical="center",
            wrap_text=False,
        )
        for cc in range(hv_panel_label_col, hv_panel_val_col + 1):
            cell = ws.cell(row=row_idx, column=cc)
            cell.fill = hv_panel_body_fill
            cell.border = thin_border
            if cc == hv_panel_val_col:
                cell.alignment = Alignment(
                    horizontal="right" if number_format else "left",
                    vertical="center",
                    wrap_text=False,
                )
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        ws.row_dimensions[row_idx].height = 24.0

    obs_lines = [str(x or "").strip() for x in hv_obs if str(x or "").strip()][:4]
    if not obs_lines:
        obs_lines = ["n/a"]
    hv_flag_source_rows = 8
    if isinstance(flags_df, pd.DataFrame) and not flags_df.empty:
        try:
            hv_flag_source_rows = max(2, int(len(flags_df.index)) + 1)
        except Exception:
            hv_flag_source_rows = 8

    def _hidden_flag_field_local(flag_row: Dict[str, Any], *names: str) -> Any:
        return hidden_flag_field(flag_row, *names)

    def _hidden_flag_score_local(value_in: Any) -> float:
        return hidden_flag_score(value_in)

    row_hv_hdr_dyn = 137
    row_hv_total_dyn = row_hv_hdr_dyn + 1
    row_hv_prof_dyn = row_hv_hdr_dyn + 2
    row_hv_cash_dyn = row_hv_hdr_dyn + 3
    row_hv_delev_dyn = row_hv_hdr_dyn + 4
    row_hv_quality_dyn = row_hv_hdr_dyn + 5
    row_hv_narr_dyn = row_hv_hdr_dyn + 6
    hidden_value_surface_model = build_hidden_value_surface_model(
        HiddenValueSurfaceModelInputs(
            flags_df=flags_df if isinstance(flags_df, pd.DataFrame) else pd.DataFrame(),
            flags_audit_df=flags_audit_df if isinstance(flags_audit_df, pd.DataFrame) else pd.DataFrame(),
            hist=hist if isinstance(hist, pd.DataFrame) else pd.DataFrame(),
            adj_metrics=adj_metrics if isinstance(adj_metrics, pd.DataFrame) else pd.DataFrame(),
            leverage_df=leverage_df if isinstance(leverage_df, pd.DataFrame) else pd.DataFrame(),
            debt_tranches=debt_tranches if isinstance(debt_tranches, pd.DataFrame) else pd.DataFrame(),
            signals_base_df=signals_base_df if isinstance(signals_base_df, pd.DataFrame) else None,
            price=price,
            build_hidden_value_flags=build_hidden_value_flags,
            build_hidden_value_flags_fallback=_build_hidden_value_flags_fallback,
            normalize_text=glx_normalize_text,
            money_formatter=_fmt_short_money_value_local,
        )
    )
    hidden_flag_rows_all = hidden_value_surface_model.rows_all
    hidden_flag_rows_triggered = hidden_value_surface_model.rows_triggered
    hidden_flag_triggered_keys = hidden_value_surface_model.triggered_keys
    hidden_flag_price_linked_keys = hidden_value_surface_model.price_linked_keys
    visible_hv_flags_hdr_row = 137
    visible_hv_flags_columns_row = visible_hv_flags_hdr_row + 1
    visible_hv_flags_start_row = visible_hv_flags_columns_row + 1
    visible_hv_flags_label_col = 1   # A
    visible_hv_flags_title_col = 2   # B
    visible_hv_flags_title_end_col = 5  # E
    visible_hv_flags_score_col = 6   # F
    visible_hv_flags_severity_col = 7  # G
    visible_hv_flags_support_col = 8  # H
    visible_hv_flags_support_end_col = 13  # M
    visible_hv_flags_panel_end_col = visible_hv_flags_support_end_col
    visible_hv_display_rows = hidden_value_surface_model.display_source_rows
    visible_hv_flag_count = hidden_value_surface_model.visible_count
    visible_hv_flags_end_row = visible_hv_flags_start_row + max(0, visible_hv_flag_count - 1)
    row_hv_obs_hdr_dyn = max(visible_hv_flags_end_row, row_hv_narr_dyn) + 2
    row_hv_obs_start_dyn = row_hv_obs_hdr_dyn + 1
    row_hv_cap_hdr_dyn = row_hv_obs_start_dyn + len(obs_lines) + 1
    row_hv_buybacks_dyn = row_hv_cap_hdr_dyn + 1
    row_hv_buybacks_note_dyn = row_hv_cap_hdr_dyn + 2
    row_hv_dividends_dyn = row_hv_cap_hdr_dyn + 4
    row_hv_dividends_note_dyn = row_hv_cap_hdr_dyn + 5

    try:
        ws.merge_cells(
            start_row=row_hv_hdr_dyn,
            start_column=hv_panel_label_col,
            end_row=row_hv_hdr_dyn,
            end_column=hv_panel_val_col,
        )
    except Exception:
        pass
    ws.cell(row=row_hv_hdr_dyn, column=hv_panel_label_col, value="Hidden Value Panel").font = bold
    for cc in range(hv_panel_label_col, hv_panel_val_col + 1):
        ws.cell(row=row_hv_hdr_dyn, column=cc).fill = section_fill
    _box_hv_panel(row_hv_total_dyn, "Hidden value score (0-100)", hv_scores.get("hidden_score"), "0")
    _box_hv_panel(row_hv_prof_dyn, "Profitability inflection", hv_scores.get("comp_profit"), "0")
    _box_hv_panel(row_hv_cash_dyn, "Cash engine", hv_scores.get("comp_cash"), "0")
    _box_hv_panel(row_hv_delev_dyn, "Deleveraging/risk down", hv_scores.get("comp_delev"), "0")
    _box_hv_panel(row_hv_quality_dyn, "Quality of earnings", hv_scores.get("comp_quality"), "0")
    _box_hv_panel(row_hv_narr_dyn, "Narrative confirmation", hv_scores.get("comp_narr"), "0")
    ws.cell(row=row_hv_obs_hdr_dyn, column=hv_label_col, value="Operating signals").font = bold
    ws.cell(row=row_hv_obs_hdr_dyn, column=hv_label_col).fill = section_fill
    ws.cell(row=row_hv_obs_hdr_dyn, column=hv_label_col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    obs_rows: List[int] = []
    for i, txt in enumerate(obs_lines, start=1):
        rr = row_hv_obs_start_dyn + (i - 1)
        _box_hv(rr, f"Signal {i}", txt)
        obs_rows.append(rr)
    ws.cell(row=row_hv_cap_hdr_dyn, column=hv_label_col, value="Capital return").font = bold
    ws.cell(row=row_hv_cap_hdr_dyn, column=hv_label_col).fill = section_fill
    ws.cell(row=row_hv_cap_hdr_dyn, column=hv_label_col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    _box_hv(row_hv_buybacks_dyn, "Buybacks (shares)", hv_buybacks)
    _box_hv(row_hv_buybacks_note_dyn, "Buybacks note", hv_buybacks_note)
    _box_hv(row_hv_dividends_dyn, "Dividends ($/share)", hv_dividends)
    _box_hv(row_hv_dividends_note_dyn, "Dividends note", hv_dividends_note)
    for rr in (row_hv_buybacks_dyn, row_hv_buybacks_note_dyn, row_hv_dividends_dyn, row_hv_dividends_note_dyn):
        ws.cell(row=rr, column=hv_label_col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for rr in obs_rows:
        ws.cell(row=rr, column=hv_label_col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        signal_cell = ws.cell(row=rr, column=hv_val_col)
        signal_cell.alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
        try:
            ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=11)  # B:K
        except Exception:
            pass
        merged_signal_width = sum(
            float(ws.column_dimensions[get_column_letter(cc)].width or 12.0)
            for cc in range(2, 12)
        )
        signal_row_h = _estimate_wrapped_row_height(
            str(signal_cell.value or ""),
            merged_signal_width,
            18.0,
            12.0,
            min_lines=1,
            max_lines=3,
        )
        ws.row_dimensions[rr].height = max(24.0, min(42.0, signal_row_h))
    for mr in list(ws.merged_cells.ranges):
        try:
            if (
                mr.min_row <= max(visible_hv_flags_end_row, row_hv_narr_dyn)
                and mr.max_row >= min(visible_hv_flags_hdr_row, row_hv_hdr_dyn)
                and mr.min_col <= visible_hv_flags_panel_end_col
                and mr.max_col >= visible_hv_flags_label_col
            ):
                ws.unmerge_cells(str(mr))
        except Exception:
            continue
    try:
        ws.merge_cells(
            start_row=visible_hv_flags_hdr_row,
            start_column=visible_hv_flags_label_col,
            end_row=visible_hv_flags_hdr_row,
            end_column=visible_hv_flags_panel_end_col,
        )
    except Exception:
        pass
    visible_hv_hdr = ws.cell(row=visible_hv_flags_hdr_row, column=visible_hv_flags_label_col, value="Hidden value flags")
    visible_hv_hdr.font = bold
    for cc in range(visible_hv_flags_label_col, visible_hv_flags_panel_end_col + 1):
        ws.cell(row=visible_hv_flags_hdr_row, column=cc).fill = section_fill
    visible_hv_hdr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    try:
        ws.merge_cells(
            start_row=row_hv_hdr_dyn,
            start_column=hv_panel_label_col,
            end_row=row_hv_hdr_dyn,
            end_column=hv_panel_val_col,
        )
    except Exception:
        pass
    ws.cell(row=row_hv_hdr_dyn, column=hv_panel_label_col, value="Hidden Value Panel").font = bold
    for cc in range(hv_panel_label_col, hv_panel_val_col + 1):
        panel_hdr_cell = ws.cell(row=row_hv_hdr_dyn, column=cc)
        panel_hdr_cell.fill = section_fill
        panel_hdr_cell.border = thin_border
        panel_hdr_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    visible_hv_header_labels = {
        visible_hv_flags_label_col: "Flag",
        visible_hv_flags_title_col: "Summary",
        visible_hv_flags_score_col: "Score",
        visible_hv_flags_severity_col: "Severity",
        visible_hv_flags_support_col: "Result / support",
    }
    for cc in range(visible_hv_flags_label_col, visible_hv_flags_panel_end_col + 1):
        header_val = visible_hv_header_labels.get(cc, "")
        header_cell = ws.cell(row=visible_hv_flags_columns_row, column=cc, value=header_val)
        header_cell.font = bold
        header_cell.fill = header_fill
        header_cell.border = thin_border
        if cc == visible_hv_flags_score_col:
            header_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        else:
            header_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    helper_col = 35  # AI
    helper_letter = get_column_letter(helper_col)
    ws.column_dimensions[helper_letter].hidden = True
    for i in range(1, visible_hv_flag_count + 1):
        rr = visible_hv_flags_start_row + (i - 1)
        no_trigger_formula = 'COUNTIF(\'Hidden_Value_Flags\'!$L:$L,">=1")=0'
        helper_formula = hidden_value_ai_helper_formula(i, rr, helper_letter)
        ws.cell(row=rr, column=helper_col, value=helper_formula)
        display_flag = visible_hv_display_rows[i - 1] if i <= len(visible_hv_display_rows) else None
        if display_flag is not None:
            display_label = f"Flag {i}"
            display_title = _hidden_flag_field_local(display_flag, "title", "Title", "flag_name", "Flag name") or ""
            display_score = _hidden_flag_score_local(_hidden_flag_field_local(display_flag, "score", "Score"))
            display_severity = _hidden_flag_field_local(display_flag, "severity", "Severity") or ""
            display_support = _hidden_flag_field_local(
                display_flag,
                "visible_support",
                "Visible support",
                "support",
                "Support",
                "evidence_1",
                "Evidence 1",
            ) or ""
        else:
            display_label = NO_TRIGGER_DISPLAY_LABEL
            display_title = NO_TRIGGER_DISPLAY_TITLE
            display_score = NO_TRIGGER_DISPLAY_SCORE
            display_severity = NO_TRIGGER_DISPLAY_SEVERITY
            display_support = NO_TRIGGER_DISPLAY_SUPPORT
        label_cell = ws.cell(
            row=rr,
            column=visible_hv_flags_label_col,
            value=display_label,
        )
        label_cell.font = bold
        try:
            ws.merge_cells(
                start_row=rr,
                start_column=visible_hv_flags_title_col,
                end_row=rr,
                end_column=visible_hv_flags_title_end_col,
            )
            ws.merge_cells(
                start_row=rr,
                start_column=visible_hv_flags_support_col,
                end_row=rr,
                end_column=visible_hv_flags_support_end_col,
            )
        except Exception:
            pass
        title_cell = ws.cell(
            row=rr,
            column=visible_hv_flags_title_col,
            value=display_title,
        )
        title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        score_cell = ws.cell(
            row=rr,
            column=visible_hv_flags_score_col,
            value=display_score,
        )
        score_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        score_cell.number_format = "0"
        severity_cell = ws.cell(
            row=rr,
            column=visible_hv_flags_severity_col,
            value=display_severity,
        )
        severity_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        support_cell = ws.cell(
            row=rr,
            column=visible_hv_flags_support_col,
            value=display_support,
        )
        support_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for cc in range(visible_hv_flags_label_col, visible_hv_flags_panel_end_col + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.border = thin_border
            if cc == visible_hv_flags_title_col:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            elif cc == visible_hv_flags_support_col:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif cc == visible_hv_flags_score_col:
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        ws.row_dimensions[rr].height = 28.0
    ws.row_dimensions[row_hv_buybacks_dyn].height = 20
    ws.cell(row=row_hv_buybacks_dyn, column=hv_val_col).alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
    try:
        ws.merge_cells(start_row=row_hv_buybacks_dyn, start_column=2, end_row=row_hv_buybacks_dyn, end_column=11)  # B:K
    except Exception:
        pass
    ws.row_dimensions[row_hv_buybacks_note_dyn].height = 20
    ws.cell(row=row_hv_buybacks_note_dyn, column=hv_val_col).alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
    try:
        ws.merge_cells(start_row=row_hv_buybacks_note_dyn, start_column=2, end_row=row_hv_buybacks_note_dyn, end_column=11)  # B:K
    except Exception:
        pass
    ws.row_dimensions[row_hv_dividends_dyn].height = 20
    ws.cell(row=row_hv_dividends_dyn, column=hv_val_col).alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
    try:
        ws.merge_cells(start_row=row_hv_dividends_dyn, start_column=2, end_row=row_hv_dividends_dyn, end_column=11)  # B:K
    except Exception:
        pass
    ws.row_dimensions[row_hv_buybacks_note_dyn].height = 34
    ws.row_dimensions[row_hv_dividends_note_dyn].height = 32
    ws.cell(row=row_hv_dividends_note_dyn, column=hv_val_col).alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
    try:
        ws.merge_cells(start_row=row_hv_dividends_note_dyn, start_column=2, end_row=row_hv_dividends_note_dyn, end_column=11)  # B:K
    except Exception:
        pass
    ws.row_dimensions[row_hv_dividends_dyn].height = 26

    valuation_export_expectation: Mapping[str, Any] | None = None
    if ctx_ref is not None:
        quarter_headers = [
            (
                _anf_visible_quarter_label(pd.Timestamp(qv).date())
                if is_anf_profile
                else f"{pd.Timestamp(qv).date().year}-Q{((pd.Timestamp(qv).date().month - 1) // 3) + 1}"
            )
            for qv in qs_ts
        ]

        def _scaled_row_values(src: Dict[pd.Timestamp, Any]) -> List[Any]:
            out_vals: List[Any] = []
            for qv in qs_ts:
                vv = src.get(pd.Timestamp(qv))
                if vv is None or (isinstance(vv, float) and pd.isna(vv)):
                    out_vals.append(None)
                else:
                    out_vals.append(float(vv) / 1e6 if isinstance(vv, (int, float)) else vv)
            return out_vals

        def _direct_row_values(src: Dict[pd.Timestamp, Any]) -> List[Any]:
            out_vals: List[Any] = []
            for qv in qs_ts:
                vv = src.get(pd.Timestamp(qv))
                out_vals.append(vv)
            return out_vals

        grid_rows_expectation: Dict[str, List[Any]] = {
            "Buybacks (cash)": _scaled_row_values(buyback_map),
            "Buybacks (TTM, cash)": _scaled_row_values(buyback_ttm_map),
            "Dividends (TTM, cash)": _scaled_row_values(dividend_ttm_map),
            "Adj EBITDA (TTM)": _scaled_row_values(adj_ebitda_ttm_map),
            "Net leverage": _direct_row_values(net_lev_display_map),
            "Cash interest coverage (TTM)": _direct_row_values(cov_cash_display_map),
        }
        if adj_ebit_ttm_map:
            grid_rows_expectation["Adj EBIT (TTM)"] = _scaled_row_values(adj_ebit_ttm_map)
        if net_lev_adj_map:
            grid_rows_expectation["Net leverage (Adj)"] = _direct_row_values(net_lev_adj_display_map)
        if cov_pnl_map:
            grid_rows_expectation["Interest coverage (P&L TTM)"] = _direct_row_values(cov_pnl_display_map)
        valuation_export_expectation = {
            "quarter_headers": quarter_headers,
            "grid_rows": grid_rows_expectation,
            "hidden_rows": {
                "Buybacks (shares)": hv_buybacks,
                "Buybacks note": hv_buybacks_note,
                "Dividends ($/share)": hv_dividends,
                "Dividends note": hv_dividends_note,
            },
        }
        ctx_ref.derived.valuation_export_expectation = valuation_export_expectation

    # Remove legacy wide Step tables from old runs so A:N stays compact.
    # Keep Hidden Value Panel + Obs + Capital return untouched.
    clear_min_row = row_hv_dividends_note_dyn + 2
    # Never clear into the valuation panel itself (starts at valuation_header_row).
    clear_max_row = min(clear_min_row + 80, valuation_header_row - 1)
    clear_min_col, clear_max_col = 1, 13  # A:M legacy cleanup area
    if clear_max_row >= clear_min_row:
        for mrange in list(ws.merged_cells.ranges):
            if (
                mrange.max_row >= clear_min_row
                and mrange.min_row <= clear_max_row
                and mrange.max_col >= clear_min_col
                and mrange.min_col <= clear_max_col
            ):
                try:
                    ws.unmerge_cells(str(mrange))
                except Exception:
                    pass
        for rr in range(clear_min_row, clear_max_row + 1):
            for cc in range(clear_min_col, clear_max_col + 1):
                c = ws.cell(row=rr, column=cc)
                c.value = None
                c.comment = None
                c.hyperlink = None
                c.fill = PatternFill(fill_type=None)
                c.border = Border()
                c.alignment = Alignment()

    return ValuationHiddenValueRenderResult(
        row_hidden_value_start=min(visible_hv_flags_hdr_row, row_hv_hdr_dyn),
        row_hidden_value_end=row_hv_dividends_note_dyn,
        row_flags_header=visible_hv_flags_hdr_row,
        row_flags_columns=visible_hv_flags_columns_row,
        row_flags_start=visible_hv_flags_start_row,
        row_flags_end=visible_hv_flags_end_row,
        row_operating_signals_header=row_hv_obs_hdr_dyn,
        row_capital_return_header=row_hv_cap_hdr_dyn,
        row_buybacks=row_hv_buybacks_dyn,
        row_buybacks_note=row_hv_buybacks_note_dyn,
        row_dividends=row_hv_dividends_dyn,
        row_dividends_note=row_hv_dividends_note_dyn,
        row_score_panel_header=row_hv_hdr_dyn,
        next_panel_row=row_hv_dividends_note_dyn + 2,
        helper_column=helper_col,
        visible_flag_count=visible_hv_flag_count,
        valuation_export_expectation=valuation_export_expectation,
    )
