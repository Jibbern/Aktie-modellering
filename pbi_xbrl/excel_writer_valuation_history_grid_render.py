"""Valuation history/source-map quarterly grid rendering support."""
from __future__ import annotations

import html
import re
import time
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, MutableMapping, Optional, Set, Tuple

import numpy as np
import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill


@dataclass(frozen=True)
class ValuationHistoryGridRenderDeps:
    runtime: MutableMapping[str, Any]


@dataclass(frozen=True)
class ValuationHistoryGridRenderResult:
    next_row: int
    valuation_row_source_values: dict[str, dict[Any, Any]]
    row_write_elapsed: float
    row_fill_elapsed: float
    _display_m_source_map_local: Any
    _margin: Any
    _ttm_map: Any
    adj_ebit_ttm_map: Any
    adj_ebitda_map: Any
    adj_ebitda_ttm_map: Any
    adj_eps_ttm_map: Any
    adj_fcf_ttm_map: Any
    ar_map: Any
    assets_map: Any
    buyback_avg_price_doc_map: Any
    buyback_cash_facts_map: Any
    buyback_doc_note_map: Any
    buyback_map: Any
    buyback_shares_map: Any
    buyback_shares_text_map: Any
    buyback_ttm_map: Any
    bv_share_map: Any
    capex_map: Any
    capex_ttm_map: Any
    capex_ttm_pct_source_map: Any
    capital_return_resolved: Any
    cash_map: Any
    cfo_map: Any
    company_operating_margin_source_map: Any
    cov_cash_display_map: Any
    cov_cash_map: Any
    cov_pnl_display_map: Any
    cov_pnl_map: Any
    debt_core_map: Any
    debt_current_map: Any
    dividend_cash_facts_map: Any
    dividend_doc_note_map: Any
    dividend_map: Any
    dividend_ps_doc_map: Any
    dividend_ttm_map: Any
    ebit_map: Any
    ebit_margin_ttm_source_map: Any
    ebitda_map: Any
    ebitda_margin_ttm_source_map: Any
    ebitda_ttm_map: Any
    fcf_conv_map: Any
    fcf_margin_ttm_source_map: Any
    fcf_per_share_ttm: Any
    fcf_ttm_map: Any
    goodwill_map: Any
    gross_profit_map: Any
    history_bv_share_source_map: Any
    history_capex_pct_source_map: Any
    history_current_ratio_source_map: Any
    history_debt_core_source_map: Any
    history_ebit_margin_source_map: Any
    history_ebitda_margin_source_map: Any
    history_eps_gaap_source_map: Any
    history_fcf_margin_source_map: Any
    history_fcf_per_share_ttm_source_map: Any
    history_fcf_source_map: Any
    history_fcf_ttm_source_map: Any
    history_gross_margin_source_map: Any
    history_net_debt_source_map: Any
    history_net_income_margin_source_map: Any
    history_owner_earnings_source_map: Any
    int_paid_ttm_map: Any
    inventory_map: Any
    last4_quarters_map: Any
    liquidity_map: Any
    net_debt_map: Any
    net_income_label: Any
    net_income_map: Any
    net_income_margin_ttm_source_map: Any
    net_income_ttm_map: Any
    net_lev_adj_display_map: Any
    net_lev_adj_map: Any
    net_lev_display_map: Any
    net_lev_map: Any
    owner_maint_capex_ratio_default: Any
    pension_map: Any
    rev_map: Any
    rev_ttm_map: Any
    row_operating_margin_pct: Any
    row_operating_margin_ttm_pct: Any
    shares_for_value_map: Any
    shares_map: Any
    shares_out_map: Any
    tbv_share_map: Any
    total_debt_map: Any
    total_equity_map: Any
    valuation_price_input_available: Any
    valuation_render_started: Any


def render_valuation_history_grid(
    deps: ValuationHistoryGridRenderDeps,
) -> ValuationHistoryGridRenderResult:
    __rt = deps.runtime
    context_globals = dict(__rt.get("context_globals") or {})

    def _rt_get(name: str) -> Any:
        if name in __rt:
            return __rt[name]
        if name in context_globals:
            return context_globals[name]
        raise KeyError(name)

    _anf_is_missing_value = _rt_get("_anf_is_missing_value")
    _anf_normalize_ytd_buyback_cash_map_for_valuation = _rt_get("_anf_normalize_ytd_buyback_cash_map_for_valuation")
    _anf_value_delta_map_for_fiscal_periods = _rt_get("_anf_value_delta_map_for_fiscal_periods")
    _anf_visible_quarter_label = _rt_get("_anf_visible_quarter_label")
    _anf_yoy_map_for_fiscal_periods = _rt_get("_anf_yoy_map_for_fiscal_periods")
    _ensure_valuation_precompute_bundle = _rt_get("_ensure_valuation_precompute_bundle")
    _ensure_valuation_render_bundle = _rt_get("_ensure_valuation_render_bundle")
    _first_existing_material_dir = _rt_get("_first_existing_material_dir")
    _operating_driver_financial_statement_files = _rt_get("_operating_driver_financial_statement_files")
    _parse_quarter_from_filename = _rt_get("_parse_quarter_from_filename")
    _parse_quarter_from_follow_text = _rt_get("_parse_quarter_from_follow_text")
    _prev_quarter_end_from_qend = _rt_get("_prev_quarter_end_from_qend")
    _quarter_notes_view = _rt_get("_quarter_notes_view")
    _read_operating_driver_text = _rt_get("_read_operating_driver_text")
    _record_writer_substage = _rt_get("_record_writer_substage")
    _resolve_col = _rt_get("_resolve_col")
    _row_fill = _rt_get("_row_fill")
    _set_cell_comment_local = _rt_get("_set_cell_comment_local")
    _timed_writer_substage = _rt_get("_timed_writer_substage")
    _valuation_row_fill_elapsed_local = _rt_get("_valuation_row_fill_elapsed_local")
    adj_metrics = _rt_get("adj_metrics")
    adj_metrics_relaxed = _rt_get("adj_metrics_relaxed")
    all_qs_ts = _rt_get("all_qs_ts")
    annual_segment_alias_patterns = _rt_get("annual_segment_alias_patterns")
    bold = _rt_get("bold")
    build_valuation_history_source_maps = _rt_get("build_valuation_history_source_maps")
    cache_dir = _rt_get("cache_dir")
    company_profile = _rt_get("company_profile")
    data_start_row = _rt_get("data_start_row")
    display_m_source_map = _rt_get("display_m_source_map")
    ew_latest_segment_financials_workbook = _rt_get("ew_latest_segment_financials_workbook")
    ew_parse_quarterly_segment_data_from_workbook = _rt_get("ew_parse_quarterly_segment_data_from_workbook")
    excel_mode = _rt_get("excel_mode")
    font_size = _rt_get("font_size")
    glx_normalize_text = _rt_get("glx_normalize_text")
    header_fill = _rt_get("header_fill")
    hist = _rt_get("hist")
    history_margin_source_map = _rt_get("history_margin_source_map")
    history_numeric_source_map = _rt_get("history_numeric_source_map")
    infer_quarter_end_from_text = _rt_get("infer_quarter_end_from_text")
    is_anf_profile = _rt_get("is_anf_profile")
    is_gpre_profile = _rt_get("is_gpre_profile")
    is_pbi_profile = _rt_get("is_pbi_profile")
    profile_ticker_txt = str(
        getattr(company_profile, "ticker", "")
        or getattr(company_profile, "symbol", "")
        or ""
    ).strip().upper()
    is_gtx_profile = profile_ticker_txt == "GTX"
    last_col = _rt_get("last_col")
    leverage_df = _rt_get("leverage_df")
    material_roots = _rt_get("material_roots")
    normalize_capex_for_valuation = _rt_get("normalize_capex_for_valuation")
    try:
        parse_adjusted_from_ex99 = _rt_get("parse_adjusted_from_ex99")
    except KeyError:
        parse_adjusted_from_ex99 = None
    price = _rt_get("price")
    qs = _rt_get("qs")
    qs_ts = _rt_get("qs_ts")
    quarter_columns = _rt_get("quarter_columns")
    quarter_key_union = _rt_get("quarter_key_union")
    quarter_notes = _rt_get("quarter_notes")
    regular_font = _rt_get("regular_font")
    source_infer_q_from_name = _rt_get("source_infer_q_from_name")
    strip_html = _rt_get("strip_html")
    style_bundle = _rt_get("style_bundle")
    thin_border = _rt_get("thin_border")
    ttm_map = _rt_get("ttm_map")
    ttm_sparse_cashflow_map = _rt_get("ttm_sparse_cashflow_map")
    valuation_soft_section_fill = _rt_get("valuation_soft_section_fill")
    ws = _rt_get("ws")

    row_write_elapsed = 0.0
    def _parse_pct_text(v: Any) -> Any:
        if not isinstance(v, str):
            return v
        s = str(v).strip()
        if not s or s.startswith("=") or "%" not in s:
            return v
        s = s.replace("\u2212", "-").replace("%", "").replace(" ", "").replace(",", ".")
        try:
            return float(s) / 100.0
        except Exception:
            return v

    def _excel_scalar_local(v: Any) -> Any:
        if v is None:
            return None
        try:
            if pd.isna(v):
                return None
        except Exception:
                pass
        return v

    # Valuation heatmap fills can legitimately compare the first visible year against
    # hidden prior history. Keep derived-row source maps broader than the visible
    # window so color logic can use real 2022 comparators for visible 2023 cells
    # without changing the saved numeric surface.
    def _quarter_key_union_local(*maps: Dict[pd.Timestamp, Any]) -> List[pd.Timestamp]:
        return quarter_key_union(all_qs_ts, *maps)

    def _set_subheader_row(row_idx: int, label: str) -> None:
        for col_idx in range(1, last_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx == 1:
                cell.value = label
            else:
                cell.value = None
            cell.fill = copy(header_fill)
            cell.border = thin_border
        if last_col > 1:
            try:
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=last_col)
            except Exception:
                pass
        sub_cell = ws.cell(row=row_idx, column=1)
        sub_cell.font = Font(
            bold=True,
            size=max(float(font_size) - 1.0, 11.0),
            color=str(style_bundle.get("text_dark") or "1F1F1F"),
        )
        sub_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_idx].height = 18.0

    valuation_row_source_values: Dict[str, Dict[pd.Timestamp, Any]] = {}

    def _set_row(row_idx: int, label: str, values: Dict[pd.Timestamp, Any], number_format: str) -> None:
        nonlocal row_write_elapsed
        row_write_started = time.perf_counter()
        normalized_source_values: Dict[pd.Timestamp, Any] = {}
        for raw_q, raw_v in dict(values or {}).items():
            q_ts = pd.to_datetime(raw_q, errors="coerce")
            if pd.isna(q_ts):
                continue
            normalized_source_values[pd.Timestamp(q_ts).normalize()] = raw_v
        # Store the full source map, not only the visible columns. The visible write
        # still uses `qs_ts`, but downstream fill logic can safely look one year back
        # into hidden history when a real comparator exists.
        valuation_row_source_values[str(label or "")] = normalized_source_values
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = regular_font
        is_pct_format = "%" in str(number_format)
        for i, _q in enumerate(qs_ts):
            col_idx = quarter_columns[i]
            val = values.get(_q)
            if is_pct_format:
                val = _parse_pct_text(val)
            val = _excel_scalar_local(val)
            data_cell = ws.cell(row=row_idx, column=col_idx, value=val)
            data_cell.number_format = number_format
        row_write_elapsed += time.perf_counter() - row_write_started

    valuation_precompute_started = time.perf_counter()

    with _timed_writer_substage("write_excel.valuation.bundle"):
        # Visible Valuation still shows the latest 12 quarters, but rolling
        # metrics need the full fiscal-history keyspace so early visible
        # periods can use hidden prior quarters (for example 2022 quarters
        # feeding visible 2023 TTM values).
        render_bundle = _ensure_valuation_render_bundle(tuple(all_qs_ts), leverage_df)
    h = render_bundle.get("hist_indexed")
    if not isinstance(h, pd.DataFrame):
        h = pd.DataFrame()
    lev = render_bundle.get("leverage_indexed")
    if not isinstance(lev, pd.DataFrame):
        lev = pd.DataFrame()
    rev_map = dict(render_bundle.get("rev_map") or {})
    gross_profit_map = dict(render_bundle.get("gross_profit_map") or {})
    ebitda_map = dict(render_bundle.get("ebitda_map") or {})
    ebit_map = dict(render_bundle.get("ebit_map") or {})
    net_income_map = dict(render_bundle.get("net_income_map") or {})
    cfo_map = dict(render_bundle.get("cfo_map") or {})
    capex_map = dict(render_bundle.get("capex_map") or {})
    price_map = dict(render_bundle.get("price_map") or {})
    market_cap_map = dict(render_bundle.get("market_cap_map") or {})
    int_paid_map = dict(render_bundle.get("int_paid_map") or {})
    tax_paid_map = dict(render_bundle.get("tax_paid_map") or {})
    cash_map = dict(render_bundle.get("cash_map") or {})
    total_debt_map = dict(render_bundle.get("total_debt_map") or {})
    debt_current_map = dict(render_bundle.get("debt_current_map") or {})
    debt_core_map = dict(render_bundle.get("debt_core_map") or {})
    shares_map = dict(render_bundle.get("shares_map") or {})
    shares_out_map = dict(render_bundle.get("shares_out_map") or {})
    total_equity_map = dict(render_bundle.get("total_equity_map") or {})
    goodwill_map = dict(render_bundle.get("goodwill_map") or {})
    intangibles_map = dict(render_bundle.get("intangibles_map") or {})
    pension_map = dict(render_bundle.get("pension_map") or {})
    assets_map = dict(render_bundle.get("assets_map") or {})
    liabilities_map = dict(render_bundle.get("liabilities_map") or {})
    assets_current_map = dict(render_bundle.get("assets_current_map") or {})
    liabilities_current_map = dict(render_bundle.get("liabilities_current_map") or {})
    ar_map = dict(render_bundle.get("ar_map") or {})
    inventory_map = dict(render_bundle.get("inventory_map") or {})
    sti_map = dict(render_bundle.get("sti_map") or {})
    rd_map = dict(render_bundle.get("rd_map") or {})
    acquisitions_map = dict(render_bundle.get("acquisitions_map") or {})
    debt_repay_map = dict(render_bundle.get("debt_repay_map") or {})
    debt_issuance_map = dict(render_bundle.get("debt_issuance_map") or {})
    ebitda_ttm_map = dict(render_bundle.get("ebitda_ttm_map") or {})
    net_lev_map = dict(render_bundle.get("net_lev_map") or {})
    cov_pnl_map = dict(render_bundle.get("cov_pnl_map") or {})
    rev_commit_map = dict(render_bundle.get("rev_commit_map") or {})
    rev_facility_map = dict(render_bundle.get("rev_facility_map") or {})
    rev_drawn_map = dict(render_bundle.get("rev_drawn_map") or {})
    rev_lc_map = dict(render_bundle.get("rev_lc_map") or {})
    rev_avail_map = dict(render_bundle.get("rev_avail_map") or {})
    liquidity_map = dict(render_bundle.get("liquidity_map") or {})
    int_paid_ttm_map = dict(render_bundle.get("int_paid_ttm_map") or {})
    buyback_map = dict(render_bundle.get("buyback_map") or {})
    dividend_map = dict(render_bundle.get("dividend_map") or {})
    buyback_cash_facts_map = dict(render_bundle.get("buyback_cash_facts_map") or {})
    dividend_cash_facts_map = dict(render_bundle.get("dividend_cash_facts_map") or {})
    buyback_shares_q_map = dict(render_bundle.get("buyback_shares_q_map") or {})
    last4_quarters_map = dict(render_bundle.get("last4_quarters_map") or {})
    valuation_precompute_bundle = _ensure_valuation_precompute_bundle(tuple(qs_ts), render_bundle)
    buyback_map = dict(valuation_precompute_bundle.get("buyback_map") or buyback_map)
    dividend_map = dict(valuation_precompute_bundle.get("dividend_map") or dividend_map)
    buyback_cash_facts_map = dict(valuation_precompute_bundle.get("buyback_cash_facts_map") or buyback_cash_facts_map)
    dividend_cash_facts_map = dict(valuation_precompute_bundle.get("dividend_cash_facts_map") or dividend_cash_facts_map)
    buyback_shares_q_map = dict(valuation_precompute_bundle.get("buyback_shares_q_map") or buyback_shares_q_map)
    buyback_shares_map = dict(valuation_precompute_bundle.get("buyback_shares_map") or {})
    buyback_shares_text_map = dict(valuation_precompute_bundle.get("buyback_shares_text_map") or {})
    buyback_avg_price_doc_map = dict(valuation_precompute_bundle.get("buyback_avg_price_doc_map") or {})
    buyback_doc_note_map = dict(valuation_precompute_bundle.get("buyback_doc_note_map") or {})
    dividend_doc_note_map = dict(valuation_precompute_bundle.get("dividend_doc_note_map") or {})
    dividend_ps_doc_map = dict(valuation_precompute_bundle.get("dividend_ps_doc_map") or {})
    valuation_audit = dict(valuation_precompute_bundle.get("valuation_audit") or {})
    capital_return_resolved = dict(valuation_precompute_bundle.get("capital_return_resolved") or {})

    def _hist_series_map_local(col_name: str) -> Dict[pd.Timestamp, Any]:
        if not isinstance(h, pd.DataFrame) or h.empty or col_name not in h.columns:
            return {}
        vals = pd.to_numeric(h.get(col_name), errors="coerce")
        out: Dict[pd.Timestamp, Any] = {}
        for idx, vv in vals.items():
            q_ts = pd.to_datetime(idx, errors="coerce")
            if pd.isna(q_ts) or pd.isna(vv):
                continue
            out[pd.Timestamp(q_ts).normalize()] = float(vv)
        return out

    marketable_securities_map = _hist_series_map_local("marketable_securities") or dict(sti_map or {})
    lease_total_map = _hist_series_map_local("lease_liabilities")
    lease_current_map = _hist_series_map_local("lease_liabilities_current")
    lease_noncurrent_map = _hist_series_map_local("lease_liabilities_noncurrent")
    if not lease_total_map and (lease_current_map or lease_noncurrent_map):
        for q in _quarter_key_union_local(lease_current_map, lease_noncurrent_map):
            lease_total_map[q] = float(lease_current_map.get(q) or 0.0) + float(lease_noncurrent_map.get(q) or 0.0)
    eps_direct_map = _hist_series_map_local("eps_diluted")

    if is_anf_profile:
        for q in _quarter_key_union_local(cash_map, debt_core_map):
            if cash_map.get(q) is not None and debt_core_map.get(q) is None:
                debt_core_map[q] = 0.0
        abl_start = pd.Timestamp(date(2026, 1, 31))
        for q in _quarter_key_union_local(cash_map, rev_commit_map, rev_facility_map, rev_drawn_map, rev_avail_map):
            q = pd.Timestamp(q).normalize()
            if q < abl_start:
                continue
            rev_commit_map[q] = 500_000_000.0
            rev_facility_map[q] = 500_000_000.0
            rev_drawn_map[q] = 0.0
            rev_lc_map[q] = 454_000.0
            rev_avail_map[q] = 449_546_000.0
            if cash_map.get(q) is not None:
                liquidity_map[q] = float(cash_map.get(q) or 0.0) + 449_546_000.0


    def _safe_float_or_none_local(value: Any) -> Optional[float]:
        try:
            coerced = pd.to_numeric(value, errors="coerce")
        except Exception:
            coerced = np.nan
        if pd.isna(coerced):
            return None
        try:
            return float(coerced)
        except Exception:
            return None

    def _yoy(
        src: Dict[pd.Timestamp, Any],
        positive_prev_only: bool = False,
        positive_cur_only: bool = False,
    ) -> Dict[pd.Timestamp, Any]:
        if is_anf_profile:
            return _anf_yoy_map_for_fiscal_periods(
                src,
                all_qs_ts,
                positive_prev_only=positive_prev_only,
                positive_cur_only=positive_cur_only,
            )
        out: Dict[pd.Timestamp, Any] = {}
        for q in _quarter_key_union_local(src):
            q = pd.Timestamp(q).normalize()
            prev = q - pd.DateOffset(years=1)
            v = _safe_float_or_none_local(src.get(q))
            p = _safe_float_or_none_local(src.get(prev))
            if v is None or p is None or p == 0:
                out[q] = None
            elif positive_prev_only and p <= 0:
                out[q] = None
            elif positive_cur_only and v <= 0:
                out[q] = None
            else:
                out[q] = (float(v) - float(p)) / abs(float(p))
        return out

    def _margin(num: Dict[pd.Timestamp, Any], denom: Dict[pd.Timestamp, Any]) -> Dict[pd.Timestamp, Any]:
        out: Dict[pd.Timestamp, Any] = {}
        for q in _quarter_key_union_local(num, denom):
            q = pd.Timestamp(q).normalize()
            n = _safe_float_or_none_local(num.get(q))
            d = _safe_float_or_none_local(denom.get(q))
            if n is None or d is None or d == 0:
                out[q] = None
            else:
                out[q] = n / d
        return out

    capex_map = normalize_capex_for_valuation(capex_map)
    fcf_map: Dict[pd.Timestamp, Any] = {}
    for q in _quarter_key_union_local(cfo_map, capex_map):
        q = pd.Timestamp(q).normalize()
        cfo = cfo_map.get(q)
        cap = capex_map.get(q)
        fcf_map[q] = (cfo - cap) if (cfo is not None and cap is not None) else None

    # Choose shares for market cap / per-share (prefer outstanding, fallback to diluted per quarter)
    shares_out_has = any(
        (shares_out_map.get(pd.Timestamp(q)) is not None) for q in qs
    ) if shares_out_map else False
    shares_for_value_map: Dict[pd.Timestamp, Any] = {}
    shares_source_map: Dict[pd.Timestamp, Optional[str]] = {}
    for q in _quarter_key_union_local(shares_out_map, shares_map):
        q = pd.Timestamp(q).normalize()
        so = _safe_float_or_none_local(shares_out_map.get(q)) if shares_out_map else None
        sd = _safe_float_or_none_local(shares_map.get(q))
        if so is not None:
            shares_for_value_map[q] = so
            shares_source_map[q] = "outstanding"
        elif sd is not None:
            shares_for_value_map[q] = sd
            shares_source_map[q] = "diluted"
        else:
            shares_for_value_map[q] = None
            shares_source_map[q] = None

    # GAAP EPS (derived from net income / diluted shares)
    eps_gaap_map: Dict[pd.Timestamp, Any] = {}
    if is_anf_profile and eps_direct_map:
        eps_gaap_map.update({pd.Timestamp(k).normalize(): v for k, v in eps_direct_map.items()})
    for q in _quarter_key_union_local(net_income_map, shares_map, eps_gaap_map):
        q = pd.Timestamp(q).normalize()
        if eps_gaap_map.get(q) is not None:
            continue
        ni = _safe_float_or_none_local(net_income_map.get(q))
        sh = _safe_float_or_none_local(shares_map.get(q))
        if ni is None or sh is None or sh == 0:
            eps_gaap_map[q] = None
        else:
            eps_gaap_map[q] = ni / sh

    # BV/TBV per share (if equity + shares exist) - prefer period-end shares outstanding
    bv_share_map: Dict[pd.Timestamp, Any] = {}
    tbv_share_map: Dict[pd.Timestamp, Any] = {}
    for q in _quarter_key_union_local(
        total_equity_map,
        shares_out_map,
        shares_for_value_map,
        goodwill_map,
        intangibles_map,
    ):
        q = pd.Timestamp(q).normalize()
        eq = _safe_float_or_none_local(total_equity_map.get(q))
        sh_out = _safe_float_or_none_local(shares_out_map.get(q))
        sh = sh_out if sh_out is not None else _safe_float_or_none_local(shares_for_value_map.get(q))
        if eq is None or sh is None or sh == 0:
            bv_share_map[q] = None
        else:
            bv_share_map[q] = eq / sh
        gw = _safe_float_or_none_local(goodwill_map.get(q))
        ia = _safe_float_or_none_local(intangibles_map.get(q))
        if eq is None or sh is None or sh == 0:
            tbv_share_map[q] = None
        else:
            if gw is None and ia is None:
                tbv_share_map[q] = None
            else:
                tbv_share_map[q] = float(eq - (gw or 0) - (ia or 0)) / sh

    def _adj_metric_value_usable_local(metric_name: str, value: Any, row_obj: Any = None) -> bool:
        vv = _safe_float_or_none_local(value)
        if vv is None:
            return False
        if is_gpre_profile:
            metric_low = str(metric_name or "").strip().lower()
            confidence = ""
            source_col = ""
            source_blob = ""
            if row_obj is not None:
                try:
                    confidence = str(row_obj.get("confidence") or "").strip().lower()
                    source_col = str(row_obj.get("col") or "").strip().lower()
                    source_blob = " ".join(
                        str(row_obj.get(name) or "")
                        for name in ("source", "source_type", "source_snippet", "doc")
                    ).lower()
                except Exception:
                    confidence = source_col = source_blob = ""
            low_conf_source = (
                "low" in confidence
                or "ocr" in source_col
                or "ocr" in source_blob
                or "transcript" in source_blob
                or "metadata" in source_blob
            )
            if metric_low in {"adj_ebit", "adj_ebitda", "adj_fcf"} and low_conf_source and 0 < abs(float(vv)) < 1_000:
                return False
            if "eps" in metric_low and low_conf_source and abs(float(vv)) >= 1.5:
                return False
        return True

    # Adjusted EBITDA from EX-99: strict first, optional relaxed (when excel_mode != clean)
    adj_ebitda_map: Dict[pd.Timestamp, Any] = {}
    adj_source_df = adj_metrics
    if (adj_source_df is None or adj_source_df.empty or "adj_ebitda" not in adj_source_df.columns) and excel_mode != "clean":
        if adj_metrics_relaxed is not None and not adj_metrics_relaxed.empty:
            adj_source_df = adj_metrics_relaxed
    if adj_source_df is not None and not adj_source_df.empty and "quarter" in adj_source_df.columns:
        am = adj_source_df.copy()
        am["quarter"] = pd.to_datetime(am["quarter"], errors="coerce")
        if "period_type" in am.columns:
            _period_order = {"annual": 0, "year": 0, "fy": 0, "ytd": 1, "quarter": 2, "": 2}
            am["_period_order"] = am["period_type"].astype(str).str.strip().str.lower().map(_period_order).fillna(2)
            am = am.sort_values(["quarter", "_period_order"], kind="stable")
        if "adj_ebitda" in am.columns:
            for _, r in am.dropna(subset=["quarter"]).iterrows():
                vv = _safe_float_or_none_local(r.get("adj_ebitda"))
                if vv is not None and _adj_metric_value_usable_local("adj_ebitda", vv, r):
                    adj_ebitda_map[pd.Timestamp(r["quarter"])] = vv

    # Adjusted FCF (from filings/slides)
    adj_fcf_map: Dict[pd.Timestamp, Any] = {}
    if adj_source_df is not None and not adj_source_df.empty and "quarter" in adj_source_df.columns:
        am = adj_source_df.copy()
        am["quarter"] = pd.to_datetime(am["quarter"], errors="coerce")
        if "period_type" in am.columns:
            _period_order = {"annual": 0, "year": 0, "fy": 0, "ytd": 1, "quarter": 2, "": 2}
            am["_period_order"] = am["period_type"].astype(str).str.strip().str.lower().map(_period_order).fillna(2)
            am = am.sort_values(["quarter", "_period_order"], kind="stable")
        if "adj_fcf" in am.columns:
            for _, r in am.dropna(subset=["quarter"]).iterrows():
                vv = _safe_float_or_none_local(r.get("adj_fcf"))
                if vv is not None and _adj_metric_value_usable_local("adj_fcf", vv, r):
                    adj_fcf_map[pd.Timestamp(r["quarter"])] = vv
    adj_ebitda_diff_map: Dict[pd.Timestamp, Any] = {}
    for q in _quarter_key_union_local(adj_ebitda_map, ebitda_map):
        q = pd.Timestamp(q).normalize()
        ae = _safe_float_or_none_local(adj_ebitda_map.get(q))
        ge = _safe_float_or_none_local(ebitda_map.get(q))
        if ae is None or ge is None:
            adj_ebitda_diff_map[q] = None
        else:
            adj_ebitda_diff_map[q] = ae - ge
    adj_fcf_diff_map: Dict[pd.Timestamp, Any] = {}
    for q in _quarter_key_union_local(adj_fcf_map, fcf_map):
        q = pd.Timestamp(q).normalize()
        af = _safe_float_or_none_local(adj_fcf_map.get(q))
        gf = _safe_float_or_none_local(fcf_map.get(q))
        if af is None or gf is None:
            adj_fcf_diff_map[q] = None
        else:
            adj_fcf_diff_map[q] = af - gf

    def _iter_adj_metric_frames_local() -> List[pd.DataFrame]:
        frames: List[pd.DataFrame] = []
        for src_df in (adj_metrics, adj_metrics_relaxed):
            if not isinstance(src_df, pd.DataFrame) or src_df.empty:
                continue
            frame = src_df.copy()
            if "quarter" not in frame.columns and "_quarter" in frame.columns:
                frame["quarter"] = frame["_quarter"]
            if "quarter" not in frame.columns:
                continue
            frame["quarter"] = pd.to_datetime(frame["quarter"], errors="coerce")
            frames.append(frame)
        return frames

    def _build_adj_metric_map_local(*candidate_cols: str) -> Dict[pd.Timestamp, Any]:
        out: Dict[pd.Timestamp, Any] = {}
        for frame in _iter_adj_metric_frames_local():
            selected_col = next((col for col in candidate_cols if col in frame.columns), None)
            if not selected_col:
                continue
            for _, r in frame.dropna(subset=["quarter"]).iterrows():
                qv = pd.Timestamp(r["quarter"])
                vv = _safe_float_or_none_local(r.get(selected_col))
                if vv is None or not _adj_metric_value_usable_local(selected_col, vv, r):
                    continue
                out[qv] = vv
        return out

    def _low_confidence_adj_metric_quarters_local(*candidate_cols: str) -> set[pd.Timestamp]:
        out: set[pd.Timestamp] = set()
        if not is_pbi_profile:
            return out
        for frame in _iter_adj_metric_frames_local():
            selected_col = next((col for col in candidate_cols if col in frame.columns), None)
            if not selected_col:
                continue
            for _, r in frame.dropna(subset=["quarter"]).iterrows():
                vv = _safe_float_or_none_local(r.get(selected_col))
                if vv is None:
                    continue
                confidence = str(r.get("confidence") or "").strip().lower()
                source_col = str(r.get("col") or "").strip().lower()
                source_note = str(r.get("source_snippet") or r.get("source") or "").strip().lower()
                if "low" in confidence or "ocr" in source_col or "ocr" in source_note:
                    out.add(pd.Timestamp(r["quarter"]).normalize())
        return out

    def _pbi_segment_adjusted_operating_maps_local() -> Tuple[Dict[pd.Timestamp, float], Dict[pd.Timestamp, float]]:
        """Fill PBI company adjusted EBIT/EBITDA gaps from source-backed segment workbook data."""
        if not is_pbi_profile:
            return {}, {}
        try:
            seg_dir = _first_existing_material_dir("segment_financials", "historical_segment")
            workbook_path = ew_latest_segment_financials_workbook(seg_dir) if seg_dir else None
            if not workbook_path:
                return {}, {}
            parsed = ew_parse_quarterly_segment_data_from_workbook(
                workbook_path,
                annual_segment_alias_patterns=annual_segment_alias_patterns,
                company_segment_alias_patterns=company_profile.segment_alias_patterns,
            )
        except Exception:
            return {}, {}
        metric_store = parsed.get("metrics") if isinstance(parsed, dict) else {}
        if not isinstance(metric_store, dict):
            return {}, {}

        def _company_component_sum(metric_name: str) -> Dict[pd.Timestamp, float]:
            metric_values = metric_store.get(metric_name)
            if not isinstance(metric_values, dict):
                return {}
            totals: Dict[pd.Timestamp, float] = {}
            counts: Dict[pd.Timestamp, int] = {}
            for seg_name, series in metric_values.items():
                seg_low = str(seg_name or "").strip().lower()
                if not seg_low or "total" in seg_low or "margin" in seg_low:
                    continue
                if not isinstance(series, dict):
                    continue
                for q_raw, v_raw in series.items():
                    vv = _safe_float_or_none_local(v_raw)
                    if vv is None:
                        continue
                    try:
                        q_ts = pd.Timestamp(q_raw).normalize()
                    except Exception:
                        continue
                    totals[q_ts] = float(totals.get(q_ts, 0.0) + vv)
                    counts[q_ts] = int(counts.get(q_ts, 0) + 1)
            return {q: total for q, total in totals.items() if counts.get(q, 0) >= 2}

        seg_adj_ebit = _company_component_sum("Adjusted EBIT")
        seg_adj_ebitda = _company_component_sum("Adjusted EBITDA")
        if not seg_adj_ebitda:
            seg_da = _company_component_sum("Depreciation & amortization")
            for q_ts, adj_ebit_v in seg_adj_ebit.items():
                da_v = _safe_float_or_none_local(seg_da.get(q_ts))
                if da_v is not None:
                    seg_adj_ebitda[q_ts] = float(adj_ebit_v + da_v)
        return seg_adj_ebit, seg_adj_ebitda

    for qv, vv in _build_adj_metric_map_local("adj_fcf").items():
        if vv is not None and adj_fcf_map.get(pd.Timestamp(qv)) is None:
            adj_fcf_map[pd.Timestamp(qv)] = vv
    for qv, vv in _build_adj_metric_map_local("adj_ebitda").items():
        q_ts = pd.Timestamp(qv)
        if vv is not None and _safe_float_or_none_local(adj_ebitda_map.get(q_ts)) is None:
            adj_ebitda_map[q_ts] = vv
    adj_ebit_map: Dict[pd.Timestamp, Any] = _build_adj_metric_map_local("adj_ebit")
    pbi_low_conf_adj_ebit_quarters = _low_confidence_adj_metric_quarters_local("adj_ebit")
    pbi_low_conf_adj_ebitda_quarters = _low_confidence_adj_metric_quarters_local("adj_ebitda")
    pbi_segment_adj_ebit_map, pbi_segment_adj_ebitda_map = _pbi_segment_adjusted_operating_maps_local()
    for qv, vv in pbi_segment_adj_ebit_map.items():
        q_ts = pd.Timestamp(qv).normalize()
        if _safe_float_or_none_local(adj_ebit_map.get(q_ts)) is None or q_ts in pbi_low_conf_adj_ebit_quarters:
            adj_ebit_map[q_ts] = vv
    for qv, vv in pbi_segment_adj_ebitda_map.items():
        q_ts = pd.Timestamp(qv).normalize()
        if _safe_float_or_none_local(adj_ebitda_map.get(q_ts)) is None or q_ts in pbi_low_conf_adj_ebitda_quarters:
            adj_ebitda_map[q_ts] = vv
    if is_anf_profile:
        for qv, vv in ebitda_map.items():
            q_ts = pd.Timestamp(qv).normalize()
            if _safe_float_or_none_local(adj_ebitda_map.get(q_ts)) is None:
                ebitda_v = _safe_float_or_none_local(vv)
                if ebitda_v is not None:
                    adj_ebitda_map[q_ts] = ebitda_v
    adj_ebitda_diff_map = {}
    for q in _quarter_key_union_local(adj_ebitda_map, ebitda_map):
        q = pd.Timestamp(q).normalize()
        ae = _safe_float_or_none_local(adj_ebitda_map.get(q))
        ge = _safe_float_or_none_local(ebitda_map.get(q))
        if ae is None or ge is None:
            adj_ebitda_diff_map[q] = None
        else:
            adj_ebitda_diff_map[q] = ae - ge
    adj_fcf_diff_map = {}
    for q in _quarter_key_union_local(adj_fcf_map, fcf_map):
        q = pd.Timestamp(q).normalize()
        af = _safe_float_or_none_local(adj_fcf_map.get(q))
        gf = _safe_float_or_none_local(fcf_map.get(q))
        if af is None or gf is None:
            adj_fcf_diff_map[q] = None
        else:
            adj_fcf_diff_map[q] = af - gf

    # Adj EPS if available (same source family as adj EBITDA)
    adj_eps_map: Dict[pd.Timestamp, Any] = _build_adj_metric_map_local(
        "adj_eps",
        "adj_eps_diluted",
        "adjusted_eps",
        "adj_eps_gaap",
    )

    def _pbi_extract_actual_adj_eps_from_text_local(text_in: Any, qd_local: Optional[date] = None) -> Optional[float]:
        txt_local = glx_normalize_text(str(text_in or ""))
        if not txt_local:
            return None
        patterns = [
            r"\badj\.?\s*eps\d*\s*\$?\(?(-?\d+\.\d{2})\)?",
            r"\badjusted\s+(?:diluted\s+)?eps\s*(?:was|of|were)?\s*\$?\(?(-?\d+\.\d{2})\)?",
            r"\badjusted\s+(?:diluted\s+)?earnings\s+per\s+(?:diluted\s+)?share\s*(?:was|of|were)?\s*\$?\(?(-?\d+\.\d{2})\)?",
        ]

        def _first_actual_from_slice(slice_txt: str) -> Optional[float]:
            for pattern in patterns:
                for match in re.finditer(pattern, slice_txt, flags=re.I):
                    window = slice_txt[max(0, match.start() - 140): match.end() + 80].lower()
                    if any(tok in window for tok in ("guidance", "outlook", "expects", "expected", "range of", "low high")):
                        continue
                    try:
                        value = float(match.group(1))
                    except Exception:
                        continue
                    if -5.0 < value < 5.0:
                        return value
            return None

        if isinstance(qd_local, date) and qd_local.month == 12:
            for marker in (
                r"fourth\s+quarter\s+20\d{2}\s+financial\s+highlights",
                r"fourth\s+quarter\s+\(\$?\s*millions",
                r"fourth\s+quarter\s+full\s+year",
            ):
                marker_match = re.search(marker, txt_local, flags=re.I)
                if marker_match:
                    q4_value = _first_actual_from_slice(txt_local[marker_match.start(): marker_match.start() + 2500])
                    if q4_value is not None:
                        return q4_value
        value = _first_actual_from_slice(txt_local)
        if value is not None:
            return value
        return None

    def _pbi_source_backed_adj_eps_map_local() -> Dict[pd.Timestamp, float]:
        if not is_pbi_profile:
            return {}
        out: Dict[pd.Timestamp, Tuple[int, float]] = {}
        rel_dirs = ("earnings_release", "CEO_letters", "earnings_transcripts")
        for root in material_roots:
            for rel_dir in rel_dirs:
                src_dir = root / rel_dir
                if not src_dir.exists() or not src_dir.is_dir():
                    continue
                try:
                    files = sorted(src_dir.iterdir(), key=lambda pp: pp.name.lower())
                except Exception:
                    continue
                for path_in in files:
                    if not path_in.is_file() or path_in.suffix.lower() not in {".htm", ".html", ".txt"}:
                        continue
                    qd_local = (
                        source_infer_q_from_name(path_in.name)
                        or _parse_quarter_from_follow_text(path_in.name)
                        or _parse_quarter_from_filename(path_in.name)
                    )
                    if not isinstance(qd_local, date):
                        continue
                    q_ts = pd.Timestamp(qd_local).normalize()
                    if q_ts.year < 2024:
                        continue
                    try:
                        raw_txt = path_in.read_text(encoding="utf-8", errors="ignore")
                    except Exception:
                        raw_txt = ""
                    if not raw_txt:
                        continue
                    plain_txt = strip_html(raw_txt) if path_in.suffix.lower() in {".htm", ".html"} else raw_txt
                    value: Optional[float] = None
                    if path_in.suffix.lower() in {".htm", ".html"}:
                        try:
                            _aebit, _aebitda, aeps, _adj, status, _col_label = parse_adjusted_from_ex99(
                                path_in.read_bytes(),
                                qd_local,
                                mode="relaxed",
                            )
                            if status in {"ok", "ok_relaxed", "ok_ocr", "ok_relaxed_ocr"}:
                                value = _safe_float_or_none_local(aeps)
                        except Exception:
                            value = None
                    if value is None:
                        value = _pbi_extract_actual_adj_eps_from_text_local(plain_txt, qd_local)
                    if value is None or not (-5.0 < float(value) < 5.0):
                        continue
                    priority = 0 if rel_dir == "earnings_release" else 1 if rel_dir == "CEO_letters" else 2
                    existing = out.get(q_ts)
                    if existing is None or priority < existing[0]:
                        out[q_ts] = (priority, float(value))
        return {q: value for q, (_priority, value) in out.items()}

    def _anf_source_backed_adj_eps_map_local() -> Dict[pd.Timestamp, float]:
        """Extract ANF quarter adjusted EPS from clean earnings-release text.

        ANF release files include compact non-GAAP EPS tables where the
        quarter-specific adjusted EPS appears before the YTD schedule.  Use
        those explicit quarter tables rather than GAAP EPS or annual/YTD
        tables so the visible quarterly model does not show false blanks.
        """
        if not is_anf_profile:
            return {}
        label_to_quarter: Dict[str, pd.Timestamp] = {}
        try:
            hist_frame = hist if isinstance(hist, pd.DataFrame) else pd.DataFrame()
        except Exception:
            hist_frame = pd.DataFrame()
        if not hist_frame.empty and "quarter" in hist_frame.columns:
            for _, hrow in hist_frame.iterrows():
                q_raw = pd.to_datetime(hrow.get("quarter"), errors="coerce")
                if pd.isna(q_raw):
                    continue
                q_ts = pd.Timestamp(q_raw).normalize()
                labels: Set[str] = set()
                visible_label = _anf_visible_quarter_label(q_ts.date())
                if visible_label:
                    labels.add(visible_label)
                fiscal_label = str(hrow.get("fiscal_label") or "").strip()
                if fiscal_label:
                    labels.add(fiscal_label)
                fy = pd.to_numeric(hrow.get("fiscal_year"), errors="coerce")
                fq = pd.to_numeric(hrow.get("fiscal_quarter"), errors="coerce")
                if pd.notna(fy) and pd.notna(fq):
                    labels.add(f"{int(fy)}-Q{int(fq)}")
                for label in labels:
                    label_to_quarter[label] = q_ts
        if not label_to_quarter:
            return {}

        quarter_words = {
            "first": 1,
            "1st": 1,
            "second": 2,
            "2nd": 2,
            "third": 3,
            "3rd": 3,
            "fourth": 4,
            "4th": 4,
        }

        def _extract_from_release_text(txt_in: str) -> Optional[Tuple[str, float]]:
            txt_norm = glx_normalize_text(txt_in)
            if not txt_norm:
                return None
            patterns = [
                r"reported\s+net\s+income\s+\(loss\)\s+per\s+diluted\s+share\s+and\s+adjusted\s+net\s+income\s+\(loss\)\s+per\s+diluted\s+share\s+for\s+the\s+(first|1st|second|2nd|third|3rd|fourth|4th)\s+quarter\s+are\s+as\s+follows:\s*(20\d{2})\s+20\d{2}.*?Adjusted\s+non-GAAP\s+\$?\s*\(?(-?\d+\.\d{2})\)?",
                r"for\s+the\s+(first|1st|second|2nd|third|3rd|fourth|4th)\s+quarter\s+are\s+as\s+follows:\s*(20\d{2})\s+20\d{2}.*?Adjusted\s+non-GAAP\s+\$?\s*\(?(-?\d+\.\d{2})\)?",
            ]
            for pattern in patterns:
                match = re.search(pattern, txt_norm, flags=re.I | re.S)
                if not match:
                    continue
                q_num = quarter_words.get(str(match.group(1)).lower())
                if not q_num:
                    continue
                try:
                    fy = int(match.group(2))
                    value = float(match.group(3))
                except Exception:
                    continue
                if -10.0 < value < 15.0:
                    return f"{fy}-Q{q_num}", value
            return None

        out: Dict[pd.Timestamp, float] = {}
        rel_dirs = ("earnings_release",)
        for root in material_roots:
            for rel_dir in rel_dirs:
                src_dir = root / rel_dir
                if not src_dir.exists() or not src_dir.is_dir():
                    continue
                try:
                    files = sorted(src_dir.iterdir(), key=lambda pp: pp.name.lower())
                except Exception:
                    continue
                for path_in in files:
                    if not path_in.is_file() or path_in.suffix.lower() not in {".htm", ".html", ".txt"}:
                        continue
                    try:
                        raw_txt = path_in.read_text(encoding="utf-8", errors="ignore")
                    except Exception:
                        raw_txt = ""
                    if not raw_txt:
                        continue
                    plain_txt = strip_html(raw_txt) if path_in.suffix.lower() in {".htm", ".html"} else raw_txt
                    plain_txt = html.unescape(plain_txt)
                    extracted = _extract_from_release_text(plain_txt)
                    if not extracted:
                        continue
                    label, value = extracted
                    q_ts = label_to_quarter.get(label)
                    if q_ts is None:
                        continue
                    out[q_ts] = float(value)
        return out

    pbi_low_conf_adj_eps_quarters = _low_confidence_adj_metric_quarters_local(
        "adj_eps",
        "adj_eps_diluted",
        "adjusted_eps",
        "adj_eps_gaap",
    )
    for qv, vv in _pbi_source_backed_adj_eps_map_local().items():
        q_ts = pd.Timestamp(qv).normalize()
        if _safe_float_or_none_local(adj_eps_map.get(q_ts)) is None or q_ts in pbi_low_conf_adj_eps_quarters:
            adj_eps_map[q_ts] = vv
    for qv, vv in _anf_source_backed_adj_eps_map_local().items():
        q_ts = pd.Timestamp(qv).normalize()
        if _safe_float_or_none_local(adj_eps_map.get(q_ts)) is None:
            adj_eps_map[q_ts] = vv

    def _ttm_map(src: Dict[pd.Timestamp, Any]) -> Dict[pd.Timestamp, Any]:
        return ttm_map(last4_quarters_map, src, all_qs_ts)

    def _complete_history_cashflow_ttm_map_local(col: str) -> Dict[pd.Timestamp, Any]:
        if hist is None or not hasattr(hist, "columns") or col not in getattr(hist, "columns", []):
            return {}
        src: Dict[pd.Timestamp, Any] = {}
        for _, row in hist.iterrows():
            try:
                q_ts = pd.Timestamp(row.get("quarter")).normalize()
            except Exception:
                continue
            if pd.isna(q_ts):
                continue
            raw = row.get(col)
            if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                src[q_ts] = None
            else:
                try:
                    src[q_ts] = float(raw)
                except Exception:
                    src[q_ts] = None
        out: Dict[pd.Timestamp, Any] = {}
        for q_now, last4_now in dict(last4_quarters_map or {}).items():
            q_key = pd.Timestamp(q_now).normalize()
            if not last4_now or len(last4_now) < 4:
                out[q_key] = None
                continue
            vals: List[float] = []
            for qv in last4_now:
                val = src.get(pd.Timestamp(qv).normalize())
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    vals = []
                    break
                vals.append(float(val))
            out[q_key] = sum(vals) if len(vals) == 4 else None
        return out

    def _history_numeric_source_map_local(col: str) -> Dict[pd.Timestamp, Any]:
        return history_numeric_source_map(hist, col)

    def _history_margin_source_map_local(num_col: str, denom_col: str = "revenue") -> Dict[pd.Timestamp, Any]:
        return history_margin_source_map(hist, num_col, denom_col, all_qs_ts)

    history_source_maps = build_valuation_history_source_maps(hist, all_qs_ts)
    history_revenue_source_map = dict(history_source_maps.get("history_revenue_source_map") or {})
    history_gross_margin_source_map = dict(history_source_maps.get("history_gross_margin_source_map") or {})
    history_ebitda_margin_source_map = dict(history_source_maps.get("history_ebitda_margin_source_map") or {})
    history_ebit_margin_source_map = dict(history_source_maps.get("history_ebit_margin_source_map") or {})
    history_net_income_margin_source_map = dict(history_source_maps.get("history_net_income_margin_source_map") or {})
    history_capex_pct_source_map = dict(history_source_maps.get("history_capex_pct_source_map") or {})
    history_fcf_source_map = dict(history_source_maps.get("history_fcf_source_map") or {})
    history_cfo_source_map = dict(history_source_maps.get("history_cfo_source_map") or {})
    history_capex_source_map = dict(history_source_maps.get("history_capex_source_map") or {})
    history_fcf_margin_source_map = dict(history_source_maps.get("history_fcf_margin_source_map") or {})
    history_owner_earnings_source_map = dict(history_source_maps.get("history_owner_earnings_source_map") or {})
    history_assets_current_source_map = dict(history_source_maps.get("history_assets_current_source_map") or {})
    history_liabilities_current_source_map = dict(history_source_maps.get("history_liabilities_current_source_map") or {})
    history_current_ratio_source_map = dict(history_source_maps.get("history_current_ratio_source_map") or {})
    history_eps_gaap_source_map = dict(history_source_maps.get("history_eps_gaap_source_map") or {})
    history_net_income_source_map = dict(history_source_maps.get("history_net_income_source_map") or {})
    history_share_denom_source_map = dict(history_source_maps.get("history_share_denom_source_map") or {})
    history_equity_source_map = dict(history_source_maps.get("history_equity_source_map") or {})
    history_shares_source_map = dict(history_source_maps.get("history_shares_source_map") or {})
    history_bv_share_source_map = dict(history_source_maps.get("history_bv_share_source_map") or {})
    history_debt_core_source_map = dict(history_source_maps.get("history_debt_core_source_map") or {})
    history_cash_source_map = dict(history_source_maps.get("history_cash_source_map") or {})
    history_net_debt_source_map = dict(history_source_maps.get("history_net_debt_source_map") or {})

    def _display_m_source_map_local(src: Dict[pd.Timestamp, Any]) -> Dict[pd.Timestamp, Any]:
        return display_m_source_map(src)

    def _ttm_sparse_cashflow_map_local(src: Dict[pd.Timestamp, Any]) -> Dict[pd.Timestamp, Any]:
        return ttm_sparse_cashflow_map(last4_quarters_map, src, all_qs_ts)

    adj_ebitda_ttm_map = _ttm_map(adj_ebitda_map) if adj_ebitda_map else {}
    adj_fcf_ttm_map = _ttm_map(adj_fcf_map) if adj_fcf_map else {}

    def _extract_quarter_safe_note_money_local(text_in: Any) -> Optional[float]:
        txt_local = glx_normalize_text(str(text_in or ""))
        if not txt_local:
            return None
        amt_match = re.search(r"\$?\s*([0-9]{1,4}(?:\.[0-9]+)?)\s*(million|m)\b", txt_local, re.I)
        if not amt_match:
            return None
        try:
            return float(amt_match.group(1)) * 1_000_000.0
        except Exception:
            return None

    def _gpre_quarter_safe_debt_overrides_local() -> Tuple[Dict[pd.Timestamp, Any], Dict[pd.Timestamp, Any]]:
        repay_overrides: Dict[pd.Timestamp, Any] = {}
        issuance_overrides: Dict[pd.Timestamp, Any] = {}
        if not is_gpre_profile:
            return repay_overrides, issuance_overrides
        note_frames_local: List[Tuple[pd.DataFrame, str]] = []
        qn_view = _quarter_notes_view(quarter_mode="date")
        if isinstance(qn_view, pd.DataFrame) and not qn_view.empty:
            note_frames_local.append((qn_view.copy(), "_quarter"))
        if isinstance(quarter_notes, pd.DataFrame) and not quarter_notes.empty:
            qn_raw = quarter_notes.copy()
            qn_quarter_col = _resolve_col(qn_raw, ["quarter", "created_quarter", "as_of_quarter"])
            if qn_quarter_col:
                qn_raw["__quarter_local__"] = pd.to_datetime(qn_raw.get(qn_quarter_col), errors="coerce").dt.date
                note_frames_local.append((qn_raw, "__quarter_local__"))

        for notes_df_local, q_col_local in note_frames_local:
            if q_col_local not in notes_df_local.columns:
                continue
            txt_cols_local = [
                cc
                for cc in [
                    _resolve_col(notes_df_local, ["note", "claim", "headline", "body"]),
                    _resolve_col(notes_df_local, ["body"]),
                    _resolve_col(notes_df_local, ["evidence_snippet", "snippet"]),
                ]
                if cc
            ]
            if not txt_cols_local:
                continue
            for _, rr_local in notes_df_local.iterrows():
                qd_local = rr_local.get(q_col_local)
                if isinstance(qd_local, datetime):
                    qd_local = qd_local.date()
                if not isinstance(qd_local, date):
                    continue
                note_txt = glx_normalize_text(
                    " ".join([str(rr_local.get(cc) or "") for cc in txt_cols_local if str(rr_local.get(cc) or "").strip()])
                )
                if not note_txt:
                    continue
                note_low = note_txt.lower()
                if (
                    "issued an additional" in note_low
                    and "convertible senior notes" in note_low
                    and "2030" in note_low
                ):
                    amt = _extract_quarter_safe_note_money_local(note_txt)
                    if amt is not None and amt > 0:
                        issuance_overrides[pd.Timestamp(qd_local)] = max(
                            float(issuance_overrides.get(pd.Timestamp(qd_local)) or 0.0),
                            float(amt),
                        )
                if (
                    "junior mezzanine debt" in note_low
                    and ("repaid" in note_low or "repay" in note_low or "eliminate" in note_low)
                    and ("obion sale proceeds" in note_low or "obion" in note_low)
                ):
                    amt = _extract_quarter_safe_note_money_local(note_txt)
                    if amt is not None and amt > 0:
                        repay_overrides[pd.Timestamp(qd_local)] = max(
                            float(repay_overrides.get(pd.Timestamp(qd_local)) or 0.0),
                            float(amt),
                        )
        local_material_records: List[Tuple[date, str, Path, str]] = []
        local_seen_paths: set[str] = set()
        candidate_rel_dirs = [
            Path("."),
            Path("materials") / "sec_primary",
            Path("materials") / "sec_exhibits",
            Path("slides_text"),
        ]
        candidate_roots = [Path(cache_dir)] + [Path(root) for root in material_roots]
        for root in candidate_roots:
            for rel_dir in candidate_rel_dirs:
                src_dir = (root / rel_dir).resolve() if rel_dir != Path(".") else root
                if not src_dir.exists() or not src_dir.is_dir():
                    continue
                try:
                    files = sorted(
                        [p for p in src_dir.iterdir() if p.is_file()],
                        key=lambda p: p.stat().st_mtime if p.exists() else 0,
                        reverse=True,
                    )[:80]
                except Exception:
                    continue
                for path_in in files:
                    if path_in.suffix.lower() not in {".pdf", ".txt", ".htm", ".html"}:
                        continue
                    name_low = path_in.name.lower()
                    if not any(tok in name_low for tok in ("2025", "q32025", "q42025", "q3-2025", "q4-2025")):
                        continue
                    try:
                        path_key = str(path_in.resolve())
                    except Exception:
                        path_key = str(path_in)
                    if path_key in local_seen_paths:
                        continue
                    joined_local = _read_operating_driver_text(path_in)
                    if not joined_local:
                        try:
                            joined_local = path_in.read_text(encoding="utf-8", errors="ignore")
                        except Exception:
                            joined_local = ""
                    if not joined_local:
                        continue
                    qd_local: Optional[date]
                    if any(tok in name_low for tok in ("q32025", "q3-2025", "20250930", "2025-09-30")):
                        qd_local = date(2025, 9, 30)
                    elif any(tok in name_low for tok in ("q42025", "q4-2025", "20251231", "2025-12-31")):
                        qd_local = date(2025, 12, 31)
                    else:
                        qd_local = (
                            _parse_quarter_from_filename(path_in.name)
                            or _parse_quarter_from_follow_text(joined_local)
                            or infer_quarter_end_from_text(joined_local)
                        )
                    if not isinstance(qd_local, date) or qd_local.year < 2025:
                        continue
                    local_seen_paths.add(path_key)
                    local_material_records.append((qd_local, "local_material", path_in, joined_local))
        for path_in in _operating_driver_financial_statement_files():
            qd_local = (
                _parse_quarter_from_filename(path_in.name)
                or _parse_quarter_from_follow_text(_read_operating_driver_text(path_in))
                or infer_quarter_end_from_text(_read_operating_driver_text(path_in))
            )
            if not isinstance(qd_local, date) or qd_local.year < 2025:
                continue
            local_material_records.append((qd_local, "financial_statement", path_in, _read_operating_driver_text(path_in)))

        for qd_local, _source_type, _path_in, joined_local in local_material_records:
            low = glx_normalize_text(joined_local).lower()
            qts = pd.Timestamp(qd_local)
            if (
                qd_local == date(2025, 9, 30)
                and "130.7" in low
                and "junior mezzanine debt" in low
                and ("eliminate" in low or "repaid" in low)
            ):
                repay_overrides[qts] = max(float(repay_overrides.get(qts) or 0.0), 130_700_000.0)
            if (
                qd_local == date(2025, 12, 31)
                and "30 million" in low
                and "2030 notes" in low
                and ("issued" in low or "subscription transactions" in low)
                and "cash" in low
            ):
                issuance_overrides[qts] = max(float(issuance_overrides.get(qts) or 0.0), 30_000_000.0)
        return repay_overrides, issuance_overrides

    def _gpre_has_explicit_q3_mezz_repay_support_local() -> bool:
        if not is_gpre_profile:
            return False
        candidate_rel_dirs = [
            Path("."),
            Path("materials") / "sec_primary",
            Path("materials") / "sec_exhibits",
            Path("slides_text"),
        ]
        seen_paths: set[str] = set()
        candidate_roots = [Path(cache_dir)] + [Path(root) for root in material_roots]
        for root in candidate_roots:
            for rel_dir in candidate_rel_dirs:
                src_dir = (root / rel_dir).resolve() if rel_dir != Path(".") else root
                if not src_dir.exists() or not src_dir.is_dir():
                    continue
                try:
                    files = sorted(
                        [p for p in src_dir.iterdir() if p.is_file()],
                        key=lambda p: p.stat().st_mtime if p.exists() else 0,
                        reverse=True,
                    )[:80]
                except Exception:
                    continue
                for path_in in files:
                    if path_in.suffix.lower() not in {".pdf", ".txt", ".htm", ".html"}:
                        continue
                    name_low = path_in.name.lower()
                    if not any(tok in name_low for tok in ("q32025", "q3-2025", "20250930", "2025-09-30", "2025")):
                        continue
                    try:
                        path_key = str(path_in.resolve())
                    except Exception:
                        path_key = str(path_in)
                    if path_key in seen_paths:
                        continue
                    seen_paths.add(path_key)
                    raw_txt = _read_operating_driver_text(path_in)
                    if not raw_txt:
                        try:
                            raw_txt = path_in.read_text(encoding="utf-8", errors="ignore")
                        except Exception:
                            raw_txt = ""
                    low = glx_normalize_text(raw_txt).lower()
                    if (
                        "130.7" in low
                        and "junior mezzanine debt" in low
                        and ("obion sale proceeds" in low or "sale of obion" in low or "obion" in low)
                        and ("repaid" in low or "repay" in low or "eliminate" in low)
                    ):
                        return True
        return False

    gpre_repay_overrides, gpre_issuance_overrides = _gpre_quarter_safe_debt_overrides_local()
    for q_override, v_override in gpre_repay_overrides.items():
        if v_override is not None:
            debt_repay_map[pd.Timestamp(q_override)] = v_override
    for q_override, v_override in gpre_issuance_overrides.items():
        if v_override is not None:
            existing_v = _safe_float_or_none_local(debt_issuance_map.get(pd.Timestamp(q_override)))
            if existing_v is None or abs(float(v_override)) > abs(float(existing_v)):
                debt_issuance_map[pd.Timestamp(q_override)] = v_override

    adj_ebit_ttm_map = _ttm_map(adj_ebit_map) if adj_ebit_map else {}
    adj_eps_ttm_map = _ttm_map(adj_eps_map) if adj_eps_map else {}
    fcf_ttm_map = _ttm_map(fcf_map)
    rev_ttm_map = _ttm_map(rev_map)
    net_income_ttm_map = _ttm_map(net_income_map) if net_income_map else {}
    ebit_ttm_map = _ttm_map(ebit_map) if ebit_map else {}
    capex_ttm_map = _ttm_map(capex_map)
    history_fcf_ttm_source_map = _ttm_map(history_fcf_source_map)
    history_fcf_per_share_ttm_source_map = (
        _margin(history_fcf_ttm_source_map, history_shares_source_map)
        if history_fcf_ttm_source_map and history_shares_source_map
        else {}
    )
    if is_anf_profile and not any(_safe_float_or_none_local(v) is not None for v in adj_fcf_map.values()):
        adj_fcf_ttm_map = dict(fcf_ttm_map)
        adj_fcf_diff_map = {
            pd.Timestamp(q).normalize(): (0.0 if v is not None else None)
            for q, v in dict(fcf_ttm_map or {}).items()
        }
    if is_anf_profile:
        buyback_map = _anf_normalize_ytd_buyback_cash_map_for_valuation(buyback_map, all_qs_ts)
        buyback_cash_facts_map = _anf_normalize_ytd_buyback_cash_map_for_valuation(buyback_cash_facts_map, all_qs_ts)
    buyback_ttm_map = dict(valuation_precompute_bundle.get("buyback_ttm_resolved_map") or (_ttm_map(buyback_map) if buyback_map else {}))
    if is_anf_profile:
        buyback_ttm_map = _ttm_sparse_cashflow_map_local(buyback_map) if buyback_map else {}
    dividend_ttm_map = dict(valuation_precompute_bundle.get("dividend_ttm_resolved_map") or (_ttm_map(dividend_map) if dividend_map else {}))
    if is_gtx_profile:
        buyback_ttm_map = _complete_history_cashflow_ttm_map_local("buybacks_cash")
        dividend_ttm_map = _complete_history_cashflow_ttm_map_local("dividends_cash")
    acquisitions_ttm_map = _ttm_map(acquisitions_map) if acquisitions_map else {}
    if is_gpre_profile:
        debt_repay_ttm_map = _ttm_sparse_cashflow_map_local(debt_repay_map) if debt_repay_map else {}
        debt_issuance_ttm_map = _ttm_sparse_cashflow_map_local(debt_issuance_map) if debt_issuance_map else {}
        if _gpre_has_explicit_q3_mezz_repay_support_local():
            repay_q = pd.Timestamp(date(2025, 9, 30))
            for q_now, last4_now in last4_quarters_map.items():
                if repay_q in tuple(pd.Timestamp(v) for v in (last4_now or ())):
                    base_val = _safe_float_or_none_local(debt_repay_ttm_map.get(pd.Timestamp(q_now))) or 0.0
                    debt_repay_ttm_map[pd.Timestamp(q_now)] = max(float(base_val), 130_700_000.0)
    else:
        debt_repay_ttm_map = _ttm_map(debt_repay_map) if debt_repay_map else {}
        debt_issuance_ttm_map = _ttm_map(debt_issuance_map) if debt_issuance_map else {}
    buyback_shares_ttm_map = dict(valuation_precompute_bundle.get("buyback_shares_ttm_resolved_map") or (_ttm_map(buyback_shares_map) if buyback_shares_map else {}))
    market_cap_for_yield_map: Dict[pd.Timestamp, Any] = {}
    for q in _quarter_key_union_local(market_cap_map, price_map, shares_for_value_map):
        q = pd.Timestamp(q).normalize()
        mc = market_cap_map.get(q)
        if mc is None:
            px = price_map.get(q)
            sh = shares_for_value_map.get(q)
            if px is not None and sh not in (None, 0):
                mc = float(px) * float(sh)
        market_cap_for_yield_map[q] = mc
    price_input_value = _safe_float_or_none_local(price)
    valuation_price_input_available = price_input_value is not None and float(price_input_value) > 0
    # Do not paint market-linked fallback formulas into the historical
    # quarterly grid when market price history is absent.  Excel can resolve
    # workbook names such as Price, but the render/readback engine used for
    # visual QA cannot, which produced visible #NAME? cells above Debt
    # Detail.  Keep scenario/price-input formulas in the lower valuation box
    # and leave historical market-linked rows blank when no clean market
    # cap exists.
    price_formula_fallback_enabled = False
    fcf_yield_ttm_map: Dict[pd.Timestamp, Any] = {}
    for q in _quarter_key_union_local(fcf_ttm_map, market_cap_for_yield_map):
        q = pd.Timestamp(q).normalize()
        fcf_t = fcf_ttm_map.get(q)
        mc = market_cap_for_yield_map.get(q)
        if fcf_t is None or mc in (None, 0) or (mc is not None and mc <= 0):
            fcf_yield_ttm_map[q] = None
        else:
            fcf_yield_ttm_map[q] = float(fcf_t) / float(mc)
    # Keep FCF yield live against Price input when market cap is missing.
    fcf_yield_ttm_display: Dict[pd.Timestamp, Any] = {}
    for q in _quarter_key_union_local(fcf_yield_ttm_map, fcf_ttm_map, shares_for_value_map, price_map):
        q = pd.Timestamp(q).normalize()
        v = fcf_yield_ttm_map.get(q)
        if v is not None:
            fcf_yield_ttm_display[q] = v
            continue
        fcf_t = fcf_ttm_map.get(q)
        sh = shares_for_value_map.get(q)
        if fcf_t is None or sh in (None, 0):
            fcf_yield_ttm_display[q] = None
            continue
        if not price_formula_fallback_enabled:
            fcf_yield_ttm_display[q] = None
            continue
        fcf_yield_ttm_display[q] = f"=IF(OR(Price=\"\",Price<=0),\"\",{float(fcf_t)}/(Price*{float(sh)}))"
    owner_maint_capex_ratio_default = 0.70
    owner_fcf_proxy_map: Dict[pd.Timestamp, Any] = {}
    for q in _quarter_key_union_local(cfo_map, capex_map):
        q = pd.Timestamp(q).normalize()
        cfo = cfo_map.get(q)
        cap = capex_map.get(q)
        if cfo is None or cap is None:
            owner_fcf_proxy_map[q] = None
        else:
            owner_fcf_proxy_map[q] = float(cfo) - float(cap) * owner_maint_capex_ratio_default

    _record_writer_substage("write_excel.valuation.precompute", valuation_precompute_started)
    valuation_render_started = time.perf_counter()

    r = data_start_row
    ws[f"A{r}"] = "Operating"
    ws[f"A{r}"].font = bold
    _row_fill(r, valuation_soft_section_fill)
    r += 1

    _set_subheader_row(r, "Top line")
    r += 1
    _set_row(r, "Revenue", {k: (v / 1e6) if v is not None else None for k, v in rev_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Revenue (TTM)", {k: (v / 1e6) if v is not None else None for k, v in rev_ttm_map.items()}, "#,##0.000")
    r += 1
    revenue_yoy_row = r
    _set_row(r, "Revenue YoY %", _yoy(rev_map), "0.0%")
    r += 1

    _set_subheader_row(r, "Margins")
    r += 1
    gross_margin_pct_map = _margin(gross_profit_map, rev_map)
    _set_row(r, "Gross margin %", gross_margin_pct_map, "0.0%")
    if history_gross_margin_source_map:
        valuation_row_source_values["Gross margin %"].update(history_gross_margin_source_map)
    r += 1
    row_operating_margin_pct = r
    company_operating_margin_map = _margin(ebit_map, rev_map)
    company_operating_margin_source_map = dict(company_operating_margin_map)
    if isinstance(hist, pd.DataFrame) and {"quarter", "revenue"}.issubset(set(hist.columns)):
        op_col = "op_income" if "op_income" in hist.columns else "operating_income" if "operating_income" in hist.columns else None
        for _idx_hist, hist_row in hist.iterrows():
            q_raw = pd.to_datetime(hist_row.get("quarter"), errors="coerce")
            if pd.isna(q_raw):
                continue
            q_key = pd.Timestamp(q_raw).normalize()
            margin_val = pd.to_numeric(hist_row.get("operating_margin"), errors="coerce")
            if pd.isna(margin_val) and op_col:
                op_val = pd.to_numeric(hist_row.get(op_col), errors="coerce")
                rev_val = pd.to_numeric(hist_row.get("revenue"), errors="coerce")
                if pd.notna(op_val) and pd.notna(rev_val) and abs(float(rev_val)) > 1e-12:
                    margin_val = float(op_val) / float(rev_val)
            if pd.notna(margin_val):
                company_operating_margin_source_map.setdefault(q_key, float(margin_val))
    _set_row(r, "Operating margin %", company_operating_margin_map, "0.0%")
    valuation_row_source_values["Operating margin %"].update(company_operating_margin_source_map)
    r += 1
    row_operating_margin_ttm_pct = r
    _set_row(r, "Operating margin (TTM)", _margin(ebit_ttm_map, rev_ttm_map), "0.0%")
    r += 1
    _set_row(r, "R&D % of revenue", _margin(rd_map, rev_map), "0.0%")
    r += 1

    _set_subheader_row(r, "Core operating")
    r += 1
    _set_row(r, "EBITDA", {k: (v / 1e6) if v is not None else None for k, v in ebitda_map.items()}, "#,##0.000")
    r += 1
    ebitda_margin_pct_map = _margin(ebitda_map, rev_map)
    _set_row(r, "EBITDA margin %", ebitda_margin_pct_map, "0.0%")
    if history_ebitda_margin_source_map:
        valuation_row_source_values["EBITDA margin %"].update(history_ebitda_margin_source_map)
    r += 1
    ebitda_yoy_row = r
    _set_row(r, "EBITDA YoY %", _yoy(ebitda_map), "0.0%")
    r += 1
    _set_row(r, "EBITDA (TTM)", {k: (v / 1e6) if v is not None else None for k, v in ebitda_ttm_map.items()}, "#,##0.000")
    r += 1
    ebitda_margin_ttm_map = _margin(ebitda_ttm_map, rev_ttm_map)
    _set_row(r, "EBITDA margin (TTM)", ebitda_margin_ttm_map, "0.0%")
    ebitda_margin_ttm_source_map = _margin(_ttm_map(_history_numeric_source_map_local("ebitda")), _ttm_map(history_revenue_source_map))
    if ebitda_margin_ttm_source_map:
        valuation_row_source_values["EBITDA margin (TTM)"].update(ebitda_margin_ttm_source_map)
    r += 1

    _set_subheader_row(r, "Adjusted operating")
    r += 1
    if is_gtx_profile:
        _set_row(r, "Adj EBIT", {k: (v / 1e6) if v is not None else None for k, v in adj_ebit_map.items()}, "#,##0.000")
        r += 1
        if adj_ebit_ttm_map or is_gtx_profile:
            _set_row(r, "Adj EBIT (TTM)", {k: (v / 1e6) if v is not None else None for k, v in adj_ebit_ttm_map.items()}, "#,##0.000")
            r += 1
        _set_row(r, "Adj EBIT margin %", _margin(adj_ebit_map, rev_map), "0.0%")
        r += 1
    _set_row(r, "Adj EBITDA", {k: (v / 1e6) if v is not None else None for k, v in adj_ebitda_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Adj EBITDA (TTM)", {k: (v / 1e6) if v is not None else None for k, v in adj_ebitda_ttm_map.items()}, "#,##0.000")
    r += 1
    if not is_gtx_profile:
        _set_row(r, "Adj EBITDA - EBITDA", {k: (v / 1e6) if v is not None else None for k, v in adj_ebitda_diff_map.items()}, "#,##0.000")
        r += 1
    adj_ebitda_margin_pct_map = _margin(adj_ebitda_map, rev_map)
    _set_row(r, "Adj EBITDA margin %", adj_ebitda_margin_pct_map, "0.0%")
    if history_ebitda_margin_source_map:
        valuation_row_source_values["Adj EBITDA margin %"].update(history_ebitda_margin_source_map)
    r += 1
    if is_gtx_profile:
        _set_row(r, "Adj FCF", {k: (v / 1e6) if v is not None else None for k, v in adj_fcf_map.items()}, "#,##0.000")
        r += 1
        adj_fcf_row = r
        _set_row(r, "Adj FCF (TTM)", {k: (v / 1e6) if v is not None else None for k, v in adj_fcf_ttm_map.items()}, "#,##0.000")
        try:
            ws.cell(row=adj_fcf_row, column=1).comment = Comment("company-defined", "Codex")
        except Exception:
            pass
        r += 1
    if is_gtx_profile:
        _set_row(r, "Adj EBITDA - EBITDA", {k: (v / 1e6) if v is not None else None for k, v in adj_ebitda_diff_map.items()}, "#,##0.000")
        r += 1
    adj_ebitda_yoy_row = r
    _set_row(r, "Adj EBITDA YoY %", _yoy(adj_ebitda_map), "0.0%")
    r += 1
    adj_ebitda_margin_ttm_map = _margin(adj_ebitda_ttm_map, rev_ttm_map)
    _set_row(r, "Adj EBITDA margin (TTM)", adj_ebitda_margin_ttm_map, "0.0%")
    if ebitda_margin_ttm_source_map:
        valuation_row_source_values["Adj EBITDA margin (TTM)"].update(ebitda_margin_ttm_source_map)
    r += 1
    if (adj_ebit_ttm_map or is_gpre_profile) and not is_gtx_profile:
        _set_row(r, "Adj EBIT (TTM)", {k: (v / 1e6) if v is not None else None for k, v in adj_ebit_ttm_map.items()}, "#,##0.000")
        r += 1

    _set_subheader_row(r, "GAAP earnings")
    r += 1
    _set_row(r, "EBIT", {k: (v / 1e6) if v is not None else None for k, v in ebit_map.items()}, "#,##0.000")
    r += 1
    ebit_margin_pct_map = _margin(ebit_map, rev_map)
    _set_row(r, "EBIT margin %", ebit_margin_pct_map, "0.0%")
    if history_ebit_margin_source_map:
        valuation_row_source_values["EBIT margin %"].update(history_ebit_margin_source_map)
    r += 1
    _set_row(r, "EBIT (TTM)", {k: (v / 1e6) if v is not None else None for k, v in ebit_ttm_map.items()}, "#,##0.000")
    r += 1
    row_ebit_margin_ttm_pct = r
    ebit_margin_ttm_map = _margin(ebit_ttm_map, rev_ttm_map)
    _set_row(r, "EBIT margin (TTM)", ebit_margin_ttm_map, "0.0%")
    ebit_margin_ttm_source_map = _margin(_ttm_map(_history_numeric_source_map_local("op_income")), _ttm_map(history_revenue_source_map))
    if ebit_margin_ttm_source_map:
        valuation_row_source_values["EBIT margin (TTM)"].update(ebit_margin_ttm_source_map)
    r += 1
    net_income_label = "Net income attrib. to A&F" if is_anf_profile else "Net income"
    _set_row(r, net_income_label, {k: (v / 1e6) if v is not None else None for k, v in net_income_map.items()}, "#,##0.000")
    r += 1
    net_income_margin_pct_map = _margin(net_income_map, rev_map)
    _set_row(r, f"{net_income_label} margin %", net_income_margin_pct_map, "0.0%")
    if history_net_income_margin_source_map:
        valuation_row_source_values[f"{net_income_label} margin %"].update(history_net_income_margin_source_map)
    r += 1
    _set_row(r, f"{net_income_label} YoY %", _yoy(net_income_map), "0.0%")
    r += 1
    _set_row(r, f"{net_income_label} (TTM)", {k: (v / 1e6) if v is not None else None for k, v in net_income_ttm_map.items()}, "#,##0.000")
    r += 1
    net_income_margin_ttm_map = _margin(net_income_ttm_map, rev_ttm_map)
    _set_row(r, f"{net_income_label} margin (TTM)", net_income_margin_ttm_map, "0.0%")
    net_income_margin_ttm_source_map = _margin(_ttm_map(_history_numeric_source_map_local("net_income")), _ttm_map(history_revenue_source_map))
    if net_income_margin_ttm_source_map:
        valuation_row_source_values[f"{net_income_label} margin (TTM)"].update(net_income_margin_ttm_source_map)
    r += 1

    ws[f"A{r}"] = "Cash Flow"
    ws[f"A{r}"].font = bold
    _row_fill(r, valuation_soft_section_fill)
    r += 1
    capex_pct_map = _margin(capex_map, rev_map)
    # TTM capex % of revenue
    capex_ttm_pct = _margin(capex_ttm_map, rev_ttm_map)
    _set_subheader_row(r, "Core cash flow")
    r += 1
    _set_row(r, "CFO", {k: (v / 1e6) if v is not None else None for k, v in cfo_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Capex", {k: (v / 1e6) if v is not None else None for k, v in capex_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Capex % of revenue", capex_pct_map, "0.0%")
    if history_capex_pct_source_map:
        valuation_row_source_values["Capex % of revenue"].update(history_capex_pct_source_map)
    r += 1
    _set_row(r, "Capex % of revenue (TTM)", capex_ttm_pct, "0.0%")
    capex_ttm_pct_source_map = _margin(_ttm_map(history_capex_source_map), _ttm_map(history_revenue_source_map))
    if capex_ttm_pct_source_map:
        valuation_row_source_values["Capex % of revenue (TTM)"].update(capex_ttm_pct_source_map)
    r += 1
    _set_row(r, "FCF (CFO-Capex)", {k: (v / 1e6) if v is not None else None for k, v in fcf_map.items()}, "#,##0.000")
    r += 1
    # FCF YoY Δ ($m) is more robust when prior-year is negative/near-zero
    if is_anf_profile:
        fcf_yoy_delta = _anf_value_delta_map_for_fiscal_periods(fcf_map, all_qs_ts, comparison="yoy")
    else:
        fcf_yoy_delta = {}
        for q in _quarter_key_union_local(fcf_map):
            q = pd.Timestamp(q).normalize()
            prev = q - pd.DateOffset(years=1)
            v = fcf_map.get(q)
            p = fcf_map.get(prev)
            if v is None or p is None:
                fcf_yoy_delta[q] = None
            else:
                fcf_yoy_delta[q] = v - p
    fcf_yoy_row = r
    _set_row(r, "FCF YoY Δ ($m)", {k: (v / 1e6) if v is not None else None for k, v in fcf_yoy_delta.items()}, "#,##0.000")
    r += 1
    _set_row(r, "FCF (TTM)", {k: (v / 1e6) if v is not None else None for k, v in fcf_ttm_map.items()}, "#,##0.000")
    r += 1
    if is_anf_profile:
        fcf_ttm_yoy_delta = _anf_value_delta_map_for_fiscal_periods(fcf_ttm_map, all_qs_ts, comparison="yoy")
    else:
        fcf_ttm_yoy_delta: Dict[pd.Timestamp, Any] = {}
        for q in _quarter_key_union_local(fcf_ttm_map):
            q = pd.Timestamp(q).normalize()
            prev = q - pd.DateOffset(years=1)
            v = fcf_ttm_map.get(q)
            p = fcf_ttm_map.get(prev)
            fcf_ttm_yoy_delta[q] = (v - p) if (v is not None and p is not None) else None
    _set_row(r, "FCF TTM YoY Δ ($m)", {k: (v / 1e6) if v is not None else None for k, v in fcf_ttm_yoy_delta.items()}, "#,##0.000")
    r += 1
    _set_subheader_row(r, "Adjusted / derived")
    r += 1
    if is_gtx_profile:
        _set_row(r, "Adj FCF - FCF", {k: (v / 1e6) if v is not None else None for k, v in adj_fcf_diff_map.items()}, "#,##0.000")
        r += 1
    else:
        _set_row(r, "Adj FCF", {k: (v / 1e6) if v is not None else None for k, v in adj_fcf_map.items()}, "#,##0.000")
        r += 1
        adj_fcf_row = r
        _set_row(r, "Adj FCF (TTM)", {k: (v / 1e6) if v is not None else None for k, v in adj_fcf_ttm_map.items()}, "#,##0.000")
        try:
            fcf_comment = "No FCF adjustment identified; Adj FCF defaults to FCF." if is_anf_profile and not adj_fcf_map else "company-defined"
            ws.cell(row=adj_fcf_row, column=1).comment = Comment(fcf_comment, "Codex")
        except Exception:
            pass
        r += 1
        _set_row(r, "Adj FCF - FCF", {k: (v / 1e6) if v is not None else None for k, v in adj_fcf_diff_map.items()}, "#,##0.000")
        r += 1
    _set_row(r, "Owner earnings (proxy)", {k: (v / 1e6) if v is not None else None for k, v in owner_fcf_proxy_map.items()}, "#,##0.000")
    r += 1
    _set_subheader_row(r, "Cash-flow quality")
    r += 1
    _set_row(r, "FCF margin %", _margin(fcf_map, rev_map), "0.0%")
    if history_fcf_margin_source_map:
        valuation_row_source_values["FCF margin %"].update(history_fcf_margin_source_map)
    r += 1
    _set_row(r, "FCF margin (TTM)", _margin(fcf_ttm_map, rev_ttm_map), "0.0%")
    fcf_margin_ttm_source_map = _margin(_ttm_map(history_fcf_source_map), _ttm_map(history_revenue_source_map))
    if fcf_margin_ttm_source_map:
        valuation_row_source_values["FCF margin (TTM)"].update(fcf_margin_ttm_source_map)
    r += 1
    _set_row(r, "Interest paid", {k: (v / 1e6) if v is not None else None for k, v in int_paid_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Tax paid", {k: (v / 1e6) if v is not None else None for k, v in tax_paid_map.items()}, "#,##0.000")
    r += 1
    _set_subheader_row(r, "Capital return / financing")
    r += 1
    _set_row(r, "Buybacks (cash)", {k: (v / 1e6) if v is not None else None for k, v in buyback_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Buybacks (TTM, cash)", {k: (v / 1e6) if v is not None else None for k, v in buyback_ttm_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Dividends (TTM, cash)", {k: (v / 1e6) if v is not None else None for k, v in dividend_ttm_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Acquisitions (TTM, cash)", {k: (v / 1e6) if v is not None else None for k, v in acquisitions_ttm_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Debt repaid (gross, TTM)", {k: (v / 1e6) if v is not None else None for k, v in debt_repay_ttm_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Debt issued (gross, TTM)", {k: (v / 1e6) if v is not None else None for k, v in debt_issuance_ttm_map.items()}, "#,##0.000")
    r += 1

    ws[f"A{r}"] = "Leverage & Liquidity"
    ws[f"A{r}"].font = bold
    _row_fill(r, valuation_soft_section_fill)
    r += 1
    _set_subheader_row(r, "Net debt position")
    r += 1
    _set_row(r, "Cash & equivalents", {k: (v / 1e6) if v is not None else None for k, v in cash_map.items()}, "#,##0.000")
    r += 1
    if marketable_securities_map:
        _set_row(r, "Marketable securities", {k: (v / 1e6) if v is not None else None for k, v in marketable_securities_map.items()}, "#,##0.000")
        r += 1
    _set_row(r, "Debt (core borrowings)", {k: (v / 1e6) if v is not None else None for k, v in debt_core_map.items()}, "#,##0.000")
    r += 1
    net_debt_map = {}
    for q in _quarter_key_union_local(debt_core_map, cash_map):
        q = pd.Timestamp(q).normalize()
        d = debt_core_map.get(q)
        c = cash_map.get(q)
        net_debt_map[q] = (d - c) if (d is not None and c is not None) else None
    for q in _quarter_key_union_local(net_debt_map, ebitda_ttm_map):
        q = pd.Timestamp(q).normalize()
        if net_lev_map.get(q) is not None:
            continue
        nd = net_debt_map.get(q)
        ebitda_ttm = ebitda_ttm_map.get(q)
        if nd is None or ebitda_ttm in (None, 0):
            continue
        try:
            ebitda_ttm_f = float(ebitda_ttm)
        except (TypeError, ValueError):
            continue
        if ebitda_ttm_f > 0:
            net_lev_map[q] = float(nd) / ebitda_ttm_f
    _set_row(r, "Net debt (core borrowings)", {k: (v / 1e6) if v is not None else None for k, v in net_debt_map.items()}, "#,##0.000")
    r += 1
    if is_anf_profile:
        net_debt_qoq_delta = _anf_value_delta_map_for_fiscal_periods(net_debt_map, all_qs_ts, comparison="qoq")
    else:
        net_debt_qoq_delta = {}
        for q in _quarter_key_union_local(net_debt_map):
            q = pd.Timestamp(q).normalize()
            prev_d = _prev_quarter_end_from_qend(q.date())
            prev = pd.Timestamp(prev_d) if prev_d else (q - pd.DateOffset(months=3))
            v = net_debt_map.get(q)
            p = net_debt_map.get(prev)
            if v is None or p is None:
                    net_debt_qoq_delta[q] = None
            else:
                net_debt_qoq_delta[q] = v - p
    _set_row(r, "Net debt QoQ Δ ($m)", {k: (v / 1e6) if v is not None else None for k, v in net_debt_qoq_delta.items()}, "#,##0.000")
    r += 1
    if is_anf_profile:
        net_debt_yoy_delta = _anf_value_delta_map_for_fiscal_periods(net_debt_map, all_qs_ts, comparison="yoy")
    else:
        net_debt_yoy_delta = {}
        for q in _quarter_key_union_local(net_debt_map):
            q = pd.Timestamp(q).normalize()
            prev = q - pd.DateOffset(years=1)
            v = net_debt_map.get(q)
            p = net_debt_map.get(prev)
            if v is None or p is None:
                net_debt_yoy_delta[q] = None
            else:
                net_debt_yoy_delta[q] = v - p
    _set_row(r, "Net debt YoY Δ ($m)", {k: (v / 1e6) if v is not None else None for k, v in net_debt_yoy_delta.items()}, "#,##0.000")
    r += 1
    if is_anf_profile:
        _set_subheader_row(r, "Supplemental net cash / lease-adjusted view")
        ws.row_dimensions[r].height = 18.0
        r += 1
        core_net_cash_map = {pd.Timestamp(k).normalize(): (-float(v) if v is not None else None) for k, v in net_debt_map.items()}
        net_cash_incl_sec_map: Dict[pd.Timestamp, Any] = {}
        lease_adj_net_debt_map: Dict[pd.Timestamp, Any] = {}
        lease_adj_net_debt_incl_sec_map: Dict[pd.Timestamp, Any] = {}
        for q in _quarter_key_union_local(debt_core_map, cash_map, marketable_securities_map, lease_total_map):
            q = pd.Timestamp(q).normalize()
            d = debt_core_map.get(q)
            c = cash_map.get(q)
            sec = marketable_securities_map.get(q) or 0.0
            leases = lease_total_map.get(q)
            net_cash_incl_sec_map[q] = (float(c) + float(sec) - float(d)) if (d is not None and c is not None) else None
            lease_adj_net_debt_map[q] = (float(d) + float(leases) - float(c)) if (d is not None and c is not None and leases is not None) else None
            lease_adj_net_debt_incl_sec_map[q] = (
                float(d) + float(leases) - float(c) - float(sec)
                if (d is not None and c is not None and leases is not None)
                else None
            )
        _set_row(r, "Core net cash", {k: (v / 1e6) if v is not None else None for k, v in core_net_cash_map.items()}, "#,##0.000")
        r += 1
        _set_row(r, "Net cash incl. securities", {k: (v / 1e6) if v is not None else None for k, v in net_cash_incl_sec_map.items()}, "#,##0.000")
        r += 1
        if lease_total_map:
            _set_row(r, "Lease liabilities", {k: (v / 1e6) if v is not None else None for k, v in lease_total_map.items()}, "#,##0.000")
            r += 1
        _set_row(r, "Lease-adjusted net debt", {k: (v / 1e6) if v is not None else None for k, v in lease_adj_net_debt_map.items()}, "#,##0.000")
        r += 1
        _set_row(r, "Lease-adjusted net debt incl. securities", {k: (v / 1e6) if v is not None else None for k, v in lease_adj_net_debt_incl_sec_map.items()}, "#,##0.000")
        r += 1
        if any(not _anf_is_missing_value(v) for v in dict(pension_map or {}).values()):
            _set_row(r, "Net pension / OPEB", {k: (v / 1e6) if v is not None else None for k, v in pension_map.items()}, "#,##0.000")
        else:
            _set_row(r, "Net pension / OPEB", {pd.Timestamp(q).normalize(): None for q in qs_ts}, "#,##0.000")
        r += 1
    elif any(not _anf_is_missing_value(v) for v in dict(pension_map or {}).values()):
        _set_row(r, "Net pension / OPEB", {k: (v / 1e6) if v is not None else None for k, v in pension_map.items()}, "#,##0.000")
        r += 1
    _set_subheader_row(r, "Coverage / leverage")
    r += 1
    _set_row(r, "EBITDA TTM", {k: (v / 1e6) if v is not None else None for k, v in ebitda_ttm_map.items()}, "#,##0.000")
    r += 1
    def _nm_display_map(raw_map: Dict[pd.Timestamp, Any], audit_key: str, denom_audit_key: str) -> Dict[pd.Timestamp, Any]:
        out: Dict[pd.Timestamp, Any] = {}
        for q in _quarter_key_union_local(raw_map):
            qts = pd.Timestamp(q).normalize()
            audit_row = dict((valuation_audit.get(qts) or {}).get(audit_key) or {})
            denom_row = dict((valuation_audit.get(qts) or {}).get(denom_audit_key) or {})
            suppress_reason = str(audit_row.get("suppress_reason") or "").strip().lower()
            if "<= 0" in suppress_reason:
                out[qts] = "N/M"
                continue
            raw_val = audit_row.get("value")
            if raw_val is None:
                raw_val = raw_map.get(qts)
            if raw_val is not None:
                out[qts] = raw_val
                continue
            denom_val = pd.to_numeric(denom_row.get("value"), errors="coerce")
            denom_val = float(denom_val) if pd.notna(denom_val) else None
            if denom_val is not None and (denom_val <= 0 or abs(denom_val) < 50_000_000.0):
                out[qts] = "N/M"
            elif (
                "missing_or_nonmeaningful_pnl_interest" in suppress_reason
                or "missing_cash_interest" in suppress_reason
            ):
                out[qts] = "N/M"
            else:
                out[qts] = None
        return out

    net_lev_display_map = _nm_display_map(net_lev_map, "net_leverage", "gaap_ebitda_ttm")
    net_lev_adj_display_map: Dict[pd.Timestamp, Any] = {}
    cov_pnl_display_map: Dict[pd.Timestamp, Any] = {}
    cov_cash_display_map: Dict[pd.Timestamp, Any] = {}
    # Net leverage (Adj EBITDA TTM) if available
    net_lev_adj_map: Dict[pd.Timestamp, Any] = {}
    if adj_ebitda_ttm_map:
        for q in _quarter_key_union_local(net_debt_map, adj_ebitda_ttm_map):
            q = pd.Timestamp(q).normalize()
            nd = net_debt_map.get(q)
            ae = adj_ebitda_ttm_map.get(q)
            if nd is None or ae in (None, 0):
                net_lev_adj_map[q] = None
            else:
                net_lev_adj_map[q] = nd / ae
    if net_lev_adj_map:
        _set_row(r, "Adj EBITDA (TTM)", {k: (v / 1e6) if v is not None else None for k, v in adj_ebitda_ttm_map.items()}, "#,##0.000")
        r += 1
    _set_row(r, "Net leverage", net_lev_display_map, "#,##0.00")
    r += 1
    if net_lev_adj_map:
        net_lev_adj_display_map = _nm_display_map(net_lev_adj_map, "net_leverage_adj", "adj_ebitda_ttm")
        _set_row(r, "Net leverage (Adj)", net_lev_adj_display_map, "#,##0.00")
        r += 1
    if cov_pnl_map:
        cov_pnl_display_map = _nm_display_map(cov_pnl_map, "interest_coverage_pnl", "gaap_ebitda_ttm")
        _set_row(r, "Interest coverage (P&L TTM)", cov_pnl_display_map, "#,##0.00")
        r += 1
    cov_cash_map = {}
    for q in _quarter_key_union_local(ebitda_ttm_map, int_paid_ttm_map):
        q = pd.Timestamp(q).normalize()
        e = ebitda_ttm_map.get(q)
        i = int_paid_ttm_map.get(q)
        if e is None or i in (None, 0):
            cov_cash_map[q] = None
        else:
            cov_cash_map[q] = e / abs(i)
    cov_cash_display_map = _nm_display_map(cov_cash_map, "cash_interest_coverage", "gaap_ebitda_ttm")
    _set_row(r, "Cash interest coverage (TTM)", cov_cash_display_map, "#,##0.00")
    r += 1
    fcf_conv_map = {}
    for q in _quarter_key_union_local(last4_quarters_map, adj_ebitda_ttm_map, ebitda_ttm_map):
        q = pd.Timestamp(q).normalize()
        e = adj_ebitda_ttm_map.get(q) if adj_ebitda_ttm_map else ebitda_ttm_map.get(q)
        # compute FCF TTM from last 4 quarters
        fcf_ttm = None
        last4 = last4_quarters_map.get(q)
        if last4:
            cfo_sum = sum([cfo_map.get(qq) or 0 for qq in last4]) if all([cfo_map.get(qq) is not None for qq in last4]) else None
            cap_sum = sum([capex_map.get(qq) or 0 for qq in last4]) if all([capex_map.get(qq) is not None for qq in last4]) else None
            if cfo_sum is not None and cap_sum is not None:
                fcf_ttm = cfo_sum - cap_sum
        if e in (None, 0) or fcf_ttm is None:
            fcf_conv_map[q] = None
        else:
            fcf_conv_map[q] = fcf_ttm / e
    _set_row(r, "FCF conversion (TTM)", fcf_conv_map, "0.0%")
    r += 1
    _set_subheader_row(r, "Revolver / liquidity")
    r += 1
    _set_row(r, "Revolver facility size", {k: (v / 1e6) if v is not None else None for k, v in rev_facility_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Revolver drawn", {k: (v / 1e6) if v is not None else None for k, v in rev_drawn_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Revolver letters of credit", {k: (v / 1e6) if v is not None else None for k, v in rev_lc_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Revolver availability", {k: (v / 1e6) if v is not None else None for k, v in rev_avail_map.items()}, "#,##0.000")
    r += 1
    show_revolver_capacity = any(
        (
            rev_commit_map.get(pd.Timestamp(qv)) is not None
            and rev_facility_map.get(pd.Timestamp(qv)) is not None
            and abs(float(rev_commit_map.get(pd.Timestamp(qv)) or 0.0) - float(rev_facility_map.get(pd.Timestamp(qv)) or 0.0)) > 1.0
        )
        for qv in qs
    )
    if show_revolver_capacity:
        _set_row(r, "Revolver capacity", {k: (v / 1e6) if v is not None else None for k, v in rev_commit_map.items()}, "#,##0.000")
        r += 1
    _set_row(r, "Liquidity (cash+availability)", {k: (v / 1e6) if v is not None else None for k, v in liquidity_map.items()}, "#,##0.000")
    r += 1
    _set_subheader_row(r, "Short-term liquidity")
    r += 1
    _set_row(r, "Current ratio", _margin(assets_current_map, liabilities_current_map), "#,##0.00")
    _set_cell_comment_local(
        ws.cell(row=r, column=1),
        "Current assets / current liabilities. Short-term liquidity measure; around 1.0+ is often healthier.",
    )
    r += 1
    quick_ratio_map: Dict[pd.Timestamp, Any] = {}
    for _q in _quarter_key_union_local(cash_map, sti_map, ar_map, liabilities_current_map):
        _q = pd.Timestamp(_q).normalize()
        cash_v = cash_map.get(_q)
        sti_v = sti_map.get(_q)
        ar_v = ar_map.get(_q)
        lc = liabilities_current_map.get(_q)
        if cash_v is None or ar_v is None or lc in (None, 0):
            quick_ratio_map[_q] = None
        else:
            quick_ratio_map[_q] = (float(cash_v) + float(sti_v or 0.0) + float(ar_v)) / float(lc)
    _set_row(r, "Quick ratio", quick_ratio_map, "#,##0.00")
    _set_cell_comment_local(
        ws.cell(row=r, column=1),
        "Near-cash current assets / current liabilities. Stricter liquidity measure; around 1.0+ is often stronger.",
    )
    r += 1

    # Equity / Per-share
    ws[f"A{r}"] = "Equity / Per-share"
    ws[f"A{r}"].font = bold
    _row_fill(r, valuation_soft_section_fill)
    r += 1
    _set_subheader_row(r, "Share count")
    r += 1
    _set_row(r, "Diluted shares (m)", {k: (v / 1e6) if v is not None else None for k, v in shares_map.items()}, "#,##0.000")
    r += 1
    _set_row(r, "Shares outstanding (m)", {k: (v / 1e6) if v is not None else None for k, v in shares_out_map.items()}, "#,##0.000")
    r += 1
    shares_change_base = shares_out_map if shares_out_has else shares_map
    shares_change_label = "Shares QoQ Δ (m) [out]" if shares_out_has else "Shares QoQ Δ (m) [dil]"
    if is_anf_profile:
        shares_qoq_delta = _anf_value_delta_map_for_fiscal_periods(shares_change_base, all_qs_ts, comparison="qoq")
    else:
        shares_qoq_delta = {}
        for q in _quarter_key_union_local(shares_change_base):
            q = pd.Timestamp(q).normalize()
            prev_d = _prev_quarter_end_from_qend(q.date())
            prev = pd.Timestamp(prev_d) if prev_d else (q - pd.DateOffset(months=3))
            v = shares_change_base.get(q)
            p = shares_change_base.get(prev)
            if v is None or p is None:
                shares_qoq_delta[q] = None
            else:
                shares_qoq_delta[q] = v - p
    _set_row(r, shares_change_label, {k: (v / 1e6) if v is not None else None for k, v in shares_qoq_delta.items()}, "#,##0.000")
    r += 1
    shares_yoy_label = "Shares YoY Δ (m) [out]" if shares_out_has else "Shares YoY Δ (m) [dil]"
    if is_anf_profile:
        shares_yoy_delta = _anf_value_delta_map_for_fiscal_periods(shares_change_base, all_qs_ts, comparison="yoy")
    else:
        shares_yoy_delta = {}
        for q in _quarter_key_union_local(shares_change_base):
            q = pd.Timestamp(q).normalize()
            prev = q - pd.DateOffset(years=1)
            v = shares_change_base.get(q)
            p = shares_change_base.get(prev)
            if v is None or p is None:
                shares_yoy_delta[q] = None
            else:
                shares_yoy_delta[q] = v - p
    _set_row(r, shares_yoy_label, {k: (v / 1e6) if v is not None else None for k, v in shares_yoy_delta.items()}, "#,##0.000")
    r += 1
    _set_subheader_row(r, "Per-share earnings")
    r += 1
    _set_row(r, "EPS (GAAP)", eps_gaap_map, "#,##0.000")
    r += 1
    if is_anf_profile:
        eps_yoy_delta = _anf_value_delta_map_for_fiscal_periods(eps_gaap_map, all_qs_ts, comparison="yoy")
    else:
        eps_yoy_delta = {}
        for q in _quarter_key_union_local(eps_gaap_map):
            q = pd.Timestamp(q).normalize()
            prev = q - pd.DateOffset(years=1)
            v = eps_gaap_map.get(q)
            p = eps_gaap_map.get(prev)
            if v is None or p is None:
                eps_yoy_delta[q] = None
            else:
                eps_yoy_delta[q] = v - p
    _set_row(r, "EPS YoY Δ ($)", eps_yoy_delta, "#,##0.000")
    r += 1
    # EPS TTM per quarter (if NI + shares available)
    eps_ttm_map: Dict[pd.Timestamp, Any] = {}
    for q in _quarter_key_union_local(last4_quarters_map, net_income_map, shares_map):
        q = pd.Timestamp(q).normalize()
        last4 = last4_quarters_map.get(q) or ()
        if len(last4) < 4:
            continue
        ni_vals = [net_income_map.get(qq) for qq in last4]
        sh_vals = [shares_map.get(qq) for qq in last4]
        if any(v is None for v in ni_vals) or any(v in (None, 0) for v in sh_vals):
            eps_ttm_map[q] = None
        else:
            ni_ttm = float(sum(ni_vals))
            sh_avg = float(sum(sh_vals)) / 4.0
            eps_ttm_map[q] = ni_ttm / sh_avg if sh_avg != 0 else None
    _set_row(r, "EPS (TTM)", eps_ttm_map, "#,##0.000")
    r += 1
    # EPS YoY % removed per request (use YoY Δ instead)
    _set_row(r, "Adj EPS", adj_eps_map or {}, "#,##0.000")
    r += 1
    _set_row(r, "Adj EPS (TTM)", adj_eps_ttm_map or {}, "#,##0.000")
    r += 1
    _set_subheader_row(r, "Per-share value")
    r += 1
    _set_row(r, "BV/share", bv_share_map, "#,##0.000")
    r += 1
    _set_row(r, "TBV/share", tbv_share_map, "#,##0.000")
    r += 1
    # FCF/share (TTM)
    fcf_per_share_ttm = {}
    for q in _quarter_key_union_local(fcf_ttm_map, shares_for_value_map):
        q = pd.Timestamp(q).normalize()
        fcf_t = fcf_ttm_map.get(q)
        sh = shares_for_value_map.get(q)
        if fcf_t is None or sh in (None, 0):
            fcf_per_share_ttm[q] = None
        else:
            fcf_per_share_ttm[q] = float(fcf_t) / float(sh)
    _set_row(r, "FCF/share (TTM)", fcf_per_share_ttm, "#,##0.000")
    r += 1
    _set_subheader_row(r, "Market-linked")
    r += 1
    # EV / yield rows (TTM-based, where market-cap and debt/cash are available)
    # If historical market cap is missing, keep rows live using current Price input
    # against quarter-specific shares/debt/cash (same pattern as FCF yield (equity)).
    ev_ttm_map: Dict[pd.Timestamp, Any] = {}
    ev_ebitda_ttm_map: Dict[pd.Timestamp, Any] = {}
    ev_adj_ebitda_ttm_map: Dict[pd.Timestamp, Any] = {}
    fcf_yield_ev_ttm_map: Dict[pd.Timestamp, Any] = {}
    ev_ttm_display: Dict[pd.Timestamp, Any] = {}
    ev_ebitda_ttm_display: Dict[pd.Timestamp, Any] = {}
    ev_adj_ebitda_ttm_display: Dict[pd.Timestamp, Any] = {}
    fcf_yield_ev_ttm_display: Dict[pd.Timestamp, Any] = {}
    for q in _quarter_key_union_local(
        market_cap_for_yield_map,
        debt_core_map,
        cash_map,
        shares_for_value_map,
        ebitda_ttm_map,
        adj_ebitda_ttm_map,
        fcf_ttm_map,
    ):
        q = pd.Timestamp(q).normalize()
        mc = market_cap_for_yield_map.get(q)
        d = debt_core_map.get(q)
        c = cash_map.get(q)
        sh = shares_for_value_map.get(q)
        ev_val = None
        if mc is not None and d is not None and c is not None:
            ev_val = float(mc) + (float(d) - float(c))
        ev_ttm_map[q] = ev_val
        eb_ttm = ebitda_ttm_map.get(q)
        aeb_ttm = adj_ebitda_ttm_map.get(q) if adj_ebitda_ttm_map else None
        fcf_t = fcf_ttm_map.get(q)
        ev_ebitda_ttm_map[q] = (ev_val / eb_ttm) if (ev_val is not None and eb_ttm not in (None, 0)) else None
        ev_adj_ebitda_ttm_map[q] = (ev_val / aeb_ttm) if (ev_val is not None and aeb_ttm not in (None, 0)) else None
        fcf_yield_ev_ttm_map[q] = (fcf_t / ev_val) if (fcf_t is not None and ev_val not in (None, 0)) else None

        # Display values: prefer numeric history, otherwise formula fallback via Price input.
        if ev_val is not None:
            ev_ttm_display[q] = float(ev_val) / 1e6
        else:
            ev_ttm_display[q] = None
        ev_ebitda_ttm_display[q] = ev_ebitda_ttm_map.get(q)
        ev_adj_ebitda_ttm_display[q] = ev_adj_ebitda_ttm_map.get(q)
        fcf_yield_ev_ttm_display[q] = fcf_yield_ev_ttm_map.get(q)

        if (
            ev_val is None
            and sh not in (None, 0)
            and d is not None
            and c is not None
            and price_formula_fallback_enabled
        ):
            ev_expr = f"(Price*{float(sh)}+{float(d)}-{float(c)})"
            ev_guard = f"OR(Price=\"\",Price<=0,({ev_expr})<=0)"
            ev_ttm_display[q] = f"=IF({ev_guard},\"\",{ev_expr}/1e6)"
            if eb_ttm not in (None, 0):
                ev_ebitda_ttm_display[q] = f"=IF({ev_guard},\"\",{ev_expr}/{float(eb_ttm)})"
            if aeb_ttm not in (None, 0):
                ev_adj_ebitda_ttm_display[q] = f"=IF({ev_guard},\"\",{ev_expr}/{float(aeb_ttm)})"
            if fcf_t is not None:
                fcf_yield_ev_ttm_display[q] = f"=IF({ev_guard},\"\",{float(fcf_t)}/{ev_expr})"

    _set_row(r, "EV ($m)", ev_ttm_display, "#,##0.000")
    r += 1
    _set_row(r, "EV/EBITDA (TTM)", ev_ebitda_ttm_display, "0.00x")
    r += 1
    _set_row(r, "EV/Adj EBITDA (TTM)", ev_adj_ebitda_ttm_display, "0.00x")
    r += 1
    _set_row(r, "FCF yield (TTM, equity)", fcf_yield_ttm_display, "0.0%")
    r += 1
    _set_row(r, "FCF yield (TTM, EV)", fcf_yield_ev_ttm_display, "0.0%")
    r += 1

    return ValuationHistoryGridRenderResult(
        next_row=r,
        valuation_row_source_values=valuation_row_source_values,
        row_write_elapsed=row_write_elapsed,
        row_fill_elapsed=float(_valuation_row_fill_elapsed_local() or 0.0),
        _display_m_source_map_local=_display_m_source_map_local,
        _margin=_margin,
        _ttm_map=_ttm_map,
        adj_ebit_ttm_map=adj_ebit_ttm_map,
        adj_ebitda_map=adj_ebitda_map,
        adj_ebitda_ttm_map=adj_ebitda_ttm_map,
        adj_eps_ttm_map=adj_eps_ttm_map,
        adj_fcf_ttm_map=adj_fcf_ttm_map,
        ar_map=ar_map,
        assets_map=assets_map,
        buyback_avg_price_doc_map=buyback_avg_price_doc_map,
        buyback_cash_facts_map=buyback_cash_facts_map,
        buyback_doc_note_map=buyback_doc_note_map,
        buyback_map=buyback_map,
        buyback_shares_map=buyback_shares_map,
        buyback_shares_text_map=buyback_shares_text_map,
        buyback_ttm_map=buyback_ttm_map,
        bv_share_map=bv_share_map,
        capex_map=capex_map,
        capex_ttm_map=capex_ttm_map,
        capex_ttm_pct_source_map=capex_ttm_pct_source_map,
        capital_return_resolved=capital_return_resolved,
        cash_map=cash_map,
        cfo_map=cfo_map,
        company_operating_margin_source_map=company_operating_margin_source_map,
        cov_cash_display_map=cov_cash_display_map,
        cov_cash_map=cov_cash_map,
        cov_pnl_display_map=cov_pnl_display_map,
        cov_pnl_map=cov_pnl_map,
        debt_core_map=debt_core_map,
        debt_current_map=debt_current_map,
        dividend_cash_facts_map=dividend_cash_facts_map,
        dividend_doc_note_map=dividend_doc_note_map,
        dividend_map=dividend_map,
        dividend_ps_doc_map=dividend_ps_doc_map,
        dividend_ttm_map=dividend_ttm_map,
        ebit_map=ebit_map,
        ebit_margin_ttm_source_map=ebit_margin_ttm_source_map,
        ebitda_map=ebitda_map,
        ebitda_margin_ttm_source_map=ebitda_margin_ttm_source_map,
        ebitda_ttm_map=ebitda_ttm_map,
        fcf_conv_map=fcf_conv_map,
        fcf_margin_ttm_source_map=fcf_margin_ttm_source_map,
        fcf_per_share_ttm=fcf_per_share_ttm,
        fcf_ttm_map=fcf_ttm_map,
        goodwill_map=goodwill_map,
        gross_profit_map=gross_profit_map,
        history_bv_share_source_map=history_bv_share_source_map,
        history_capex_pct_source_map=history_capex_pct_source_map,
        history_current_ratio_source_map=history_current_ratio_source_map,
        history_debt_core_source_map=history_debt_core_source_map,
        history_ebit_margin_source_map=history_ebit_margin_source_map,
        history_ebitda_margin_source_map=history_ebitda_margin_source_map,
        history_eps_gaap_source_map=history_eps_gaap_source_map,
        history_fcf_margin_source_map=history_fcf_margin_source_map,
        history_fcf_per_share_ttm_source_map=history_fcf_per_share_ttm_source_map,
        history_fcf_source_map=history_fcf_source_map,
        history_fcf_ttm_source_map=history_fcf_ttm_source_map,
        history_gross_margin_source_map=history_gross_margin_source_map,
        history_net_debt_source_map=history_net_debt_source_map,
        history_net_income_margin_source_map=history_net_income_margin_source_map,
        history_owner_earnings_source_map=history_owner_earnings_source_map,
        int_paid_ttm_map=int_paid_ttm_map,
        inventory_map=inventory_map,
        last4_quarters_map=last4_quarters_map,
        liquidity_map=liquidity_map,
        net_debt_map=net_debt_map,
        net_income_label=net_income_label,
        net_income_map=net_income_map,
        net_income_margin_ttm_source_map=net_income_margin_ttm_source_map,
        net_income_ttm_map=net_income_ttm_map,
        net_lev_adj_display_map=net_lev_adj_display_map,
        net_lev_adj_map=net_lev_adj_map,
        net_lev_display_map=net_lev_display_map,
        net_lev_map=net_lev_map,
        owner_maint_capex_ratio_default=owner_maint_capex_ratio_default,
        pension_map=pension_map,
        rev_map=rev_map,
        rev_ttm_map=rev_ttm_map,
        row_operating_margin_pct=row_operating_margin_pct,
        row_operating_margin_ttm_pct=row_operating_margin_ttm_pct,
        shares_for_value_map=shares_for_value_map,
        shares_map=shares_map,
        shares_out_map=shares_out_map,
        tbv_share_map=tbv_share_map,
        total_debt_map=total_debt_map,
        total_equity_map=total_equity_map,
        valuation_price_input_available=valuation_price_input_available,
        valuation_render_started=valuation_render_started
    )
