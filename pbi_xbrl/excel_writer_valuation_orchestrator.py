"""Valuation sheet orchestrator.

This module owns the remaining Valuation sheet orchestration after the source,
history-grid, formula-core, Debt Detail, Hidden Value, guidance, side-panel,
trend, sensitivity, and final-layout helpers have been extracted. The writer
context keeps only the callback wrapper and shared helper ownership.
"""
from __future__ import annotations

import copy
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Mapping, MutableMapping, Optional

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from .excel_writer_valuation_debt_detail_render import (
    ValuationDebtDetailRenderDeps,
    render_valuation_debt_detail,
)
from .excel_writer_valuation_final_layout import (
    ValuationFinalLayoutDeps,
    apply_valuation_final_layout,
)
from .excel_writer_valuation_formula_core_render import (
    ValuationFormulaCoreRenderDeps,
    render_valuation_formula_core,
)
from .excel_writer_valuation_guidance_render import (
    ValuationGuidanceRenderDeps,
    render_valuation_guidance_panel,
)
from .excel_writer_valuation_hidden_value_render import (
    ValuationHiddenValueRenderDeps,
    render_valuation_hidden_value_panel,
)
from .excel_writer_valuation_hidden_value_state import (
    ValuationHiddenValueStateDeps,
    build_valuation_hidden_value_state,
)
from .excel_writer_valuation_history_grid_render import (
    ValuationHistoryGridRenderDeps,
    render_valuation_history_grid,
)
from .excel_writer_valuation_operating_thesis_render import (
    ValuationOperatingThesisRenderDeps,
    render_valuation_operating_thesis_panels,
)
from .excel_writer_valuation_sensitivity_heatmap_render import (
    ValuationSensitivityHeatmapRenderDeps,
    render_valuation_sensitivity_heatmaps,
)
from .excel_writer_valuation_trend_flags_render import (
    ValuationTrendFlagsRenderDeps,
    render_valuation_trend_flags_panel,
)


@dataclass(frozen=True)
class ValuationOrchestratorDeps:
    wb: Any
    ticker: str
    company_profile: Any
    is_pbi_profile: bool
    is_gpre_profile: bool
    is_anf_profile: bool
    price: Any
    excel_mode: Any
    hist: Any
    quarter_notes: Any
    promises: Any
    audit: Any
    promise_progress: Any
    slides_guidance: Any
    slides_debt: Any
    valuation_grid_df: Any
    adj_metrics: Any
    adj_metrics_relaxed: Any
    leverage_df: Any
    manifest_df: Any
    flags_df: Any
    flags_audit_df: Any
    signals_base_df: Any
    debt_tranches: Any
    debt_tranches_latest: Any
    debt_credit_notes: Any
    company_overview: Any
    cache_root: Any
    cache_dir: Any
    material_roots: Any
    ctx_ref: Any
    ui_state: MutableMapping[str, Any]
    context_globals: MutableMapping[str, Any]
    get_valuation_style_bundle: Callable[..., Mapping[str, Any]]
    set_cell_comment: Callable[..., None]
    timed_writer_substage: Callable[..., Any]
    record_writer_substage: Callable[..., None]
    record_writer_elapsed: Callable[..., None]
    context_helpers: Mapping[str, Any]


def write_valuation_sheet(deps: ValuationOrchestratorDeps) -> None:
    wb = deps.wb
    ticker = deps.ticker
    company_profile = deps.company_profile
    is_pbi_profile = deps.is_pbi_profile
    is_gpre_profile = deps.is_gpre_profile
    is_anf_profile = deps.is_anf_profile
    price = deps.price
    excel_mode = deps.excel_mode
    hist = deps.hist
    quarter_notes = deps.quarter_notes
    promises = deps.promises
    audit = deps.audit
    promise_progress = deps.promise_progress
    slides_guidance = deps.slides_guidance
    slides_debt = deps.slides_debt
    valuation_grid_df = deps.valuation_grid_df
    adj_metrics = deps.adj_metrics
    adj_metrics_relaxed = deps.adj_metrics_relaxed
    leverage_df = deps.leverage_df
    manifest_df = deps.manifest_df
    flags_df = deps.flags_df
    flags_audit_df = deps.flags_audit_df
    signals_base_df = deps.signals_base_df
    debt_tranches = deps.debt_tranches
    debt_tranches_latest = deps.debt_tranches_latest
    debt_credit_notes = deps.debt_credit_notes
    company_overview = deps.company_overview
    cache_root = deps.cache_root
    cache_dir = deps.cache_dir
    material_roots = deps.material_roots
    ctx_ref = deps.ctx_ref
    ui_state = deps.ui_state
    context_globals = deps.context_globals
    _get_valuation_style_bundle = deps.get_valuation_style_bundle
    _set_cell_comment_local = deps.set_cell_comment
    _timed_writer_substage = deps.timed_writer_substage
    _record_writer_substage = deps.record_writer_substage
    _record_writer_elapsed = deps.record_writer_elapsed
    runtime_globals = {**context_globals, **globals(), **dict(deps.context_helpers)}
    globals().update(runtime_globals)

    ws = wb.create_sheet(title="Valuation")
    ws.sheet_format.defaultRowHeight = 18
    ws.sheet_view.zoomScale = 110
    style_bundle = _get_valuation_style_bundle()
    # Keep quarter header close to top and freeze at the data start row.
    actuals_row = 5
    quarter_row = 6
    data_start_row = 7
    ws.freeze_panes = f"B{data_start_row}"

    def _px_to_width(px: float) -> float:
        # Approx conversion: Excel column width units from pixels.
        try:
            p = float(px)
        except Exception:
            p = 100.0
        return max(1.0, round((p - 5.0) / 7.0, 2))

    ws["A1"] = "Scale"
    ws["B1"] = "$m"
    ws["A2"] = "Values scaled to $m unless %"
    try:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=13)  # A:M
    except Exception:
        pass
    ws["A3"] = "Valuation"
    ws["A3"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"].fill = copy(style_bundle["title_fill"])
    for cc in range(1, 14):
        ws.cell(row=3, column=cc).fill = copy(style_bundle["title_fill"])
    legend = [
        (3, "<=-15%", "A63A00"),
        (4, "-15..-5", "D55E00"),
        (5, "-5..+5", "DDDDDD"),
        (6, "+5..+15", "9BD3F5"),
        (7, ">=+15%", "2F80ED"),
    ]
    for col, txt, color in legend:
        c = ws.cell(row=1, column=col, value=txt)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")

    if hist is None or hist.empty or "quarter" not in hist.columns:
        ws["A4"] = "No data."
        return

    all_qs = sorted(pd.to_datetime(hist["quarter"], errors="coerce").dropna().unique())
    if not all_qs:
        ws["A4"] = "No data."
        return
    qs = all_qs[-12:]
    qs_ts = [pd.Timestamp(q) for q in qs]
    all_qs_ts = [pd.Timestamp(q).normalize() for q in all_qs]
    start_col = 2  # B
    last_col = start_col + len(qs) - 1
    last_col_letter = get_column_letter(last_col)
    quarter_columns = [start_col + i for i in range(len(qs_ts))]

    header_fill = copy(style_bundle["header_fill"])
    section_fill = copy(style_bundle["section_fill"])
    valuation_soft_section_fill = copy(style_bundle.get("valuation_soft_section_fill") or style_bundle["section_fill"])
    title_fill = copy(style_bundle["title_fill"])
    input_fill = copy(style_bundle["input_fill"])
    bold = copy(style_bundle["bold_font"])
    regular_font = Font(size=font_size, bold=False)
    center_alignment = Alignment(horizontal="center")
    left_alignment = Alignment(horizontal="left")
    thick = Side(style="thick")
    thin_border = copy(style_bundle["thin_border"])

    # Header rows start after guidance canvas.
    for i, q in enumerate(qs_ts):
        col_idx = quarter_columns[i]
        qd = q.date()
        q_label = _anf_visible_quarter_label(qd) if is_anf_profile else f"{qd.year}-Q{((qd.month - 1) // 3) + 1}"
        quarter_cell = ws.cell(row=quarter_row, column=col_idx, value=q_label)
        quarter_cell.font = bold
        quarter_cell.alignment = center_alignment
        quarter_cell.fill = header_fill
        quarter_cell.border = thin_border

    # header rows
    quarter_hdr_cell = ws.cell(row=quarter_row, column=1, value="Quarter")
    quarter_hdr_cell.font = bold
    quarter_hdr_cell.alignment = left_alignment
    quarter_hdr_cell.fill = header_fill
    quarter_hdr_cell.border = thin_border
    ws.merge_cells(start_row=actuals_row, start_column=start_col, end_row=actuals_row, end_column=last_col)
    actuals_cell = ws.cell(row=actuals_row, column=start_col, value="Actuals")
    actuals_cell.font = bold
    actuals_cell.alignment = center_alignment
    actuals_cell.fill = header_fill
    actuals_cell.border = thin_border
    for col in quarter_columns:
        ws.cell(row=actuals_row, column=col).fill = header_fill
        ws.cell(row=actuals_row, column=col).border = thin_border
        ws.cell(row=quarter_row, column=col).fill = header_fill
        ws.cell(row=quarter_row, column=col).border = thin_border

    row_fill_elapsed = 0.0
    row_write_elapsed = 0.0

    def _row_fill(row_idx: int, fill: PatternFill) -> None:
        nonlocal row_fill_elapsed
        row_fill_started = time.perf_counter()
        for col in range(1, last_col + 1):
            ws.cell(row=row_idx, column=col).fill = fill
        row_fill_elapsed += time.perf_counter() - row_fill_started

    def _valuation_row_fill_elapsed_local() -> float:
        return float(row_fill_elapsed or 0.0)

    history_grid_result = render_valuation_history_grid(
        ValuationHistoryGridRenderDeps(
            runtime={
                **runtime_globals,
                **locals(),
                "context_globals": context_globals,
                "_valuation_row_fill_elapsed_local": _valuation_row_fill_elapsed_local,
                "_anf_is_missing_value": _anf_is_missing_value,
                "_anf_normalize_ytd_buyback_cash_map_for_valuation": _anf_normalize_ytd_buyback_cash_map_for_valuation,
                "_anf_value_delta_map_for_fiscal_periods": _anf_value_delta_map_for_fiscal_periods,
                "_anf_visible_quarter_label": _anf_visible_quarter_label,
                "_anf_yoy_map_for_fiscal_periods": _anf_yoy_map_for_fiscal_periods,
                "_ensure_valuation_precompute_bundle": _ensure_valuation_precompute_bundle,
                "_ensure_valuation_render_bundle": _ensure_valuation_render_bundle,
                "_first_existing_material_dir": _first_existing_material_dir,
                "_operating_driver_financial_statement_files": _operating_driver_financial_statement_files,
                "_parse_quarter_from_filename": _parse_quarter_from_filename,
                "_parse_quarter_from_follow_text": _parse_quarter_from_follow_text,
                "_prev_quarter_end_from_qend": _prev_quarter_end_from_qend,
                "_quarter_notes_view": _quarter_notes_view,
                "_read_operating_driver_text": _read_operating_driver_text,
                "_record_writer_substage": _record_writer_substage,
                "_resolve_col": _resolve_col,
                "_row_fill": _row_fill,
                "_set_cell_comment_local": _set_cell_comment_local,
                "_timed_writer_substage": _timed_writer_substage,
                "adj_metrics": adj_metrics,
                "adj_metrics_relaxed": adj_metrics_relaxed,
                "all_qs_ts": all_qs_ts,
                "annual_segment_alias_patterns": annual_segment_alias_patterns,
                "bold": bold,
                "build_valuation_history_source_maps": build_valuation_history_source_maps,
                "cache_dir": cache_dir,
                "company_profile": company_profile,
                "data_start_row": data_start_row,
                "display_m_source_map": display_m_source_map,
                "ew_latest_segment_financials_workbook": ew_latest_segment_financials_workbook,
                "ew_parse_quarterly_segment_data_from_workbook": ew_parse_quarterly_segment_data_from_workbook,
                "excel_mode": excel_mode,
                "font_size": font_size,
                "glx_normalize_text": glx_normalize_text,
                "header_fill": header_fill,
                "hist": hist,
                "history_margin_source_map": history_margin_source_map,
                "history_numeric_source_map": history_numeric_source_map,
                "infer_quarter_end_from_text": infer_quarter_end_from_text,
                "is_anf_profile": is_anf_profile,
                "is_gpre_profile": is_gpre_profile,
                "is_pbi_profile": is_pbi_profile,
                "last_col": last_col,
                "leverage_df": leverage_df,
                "material_roots": material_roots,
                "normalize_capex_for_valuation": normalize_capex_for_valuation,
                "price": price,
                "qs": qs,
                "qs_ts": qs_ts,
                "quarter_columns": quarter_columns,
                "quarter_key_union": quarter_key_union,
                "quarter_notes": quarter_notes,
                "regular_font": regular_font,
                "source_infer_q_from_name": source_infer_q_from_name,
                "strip_html": strip_html,
                "style_bundle": style_bundle,
                "thin_border": thin_border,
                "ttm_map": ttm_map,
                "ttm_sparse_cashflow_map": ttm_sparse_cashflow_map,
                "valuation_soft_section_fill": valuation_soft_section_fill,
                "ws": ws,
            }
        )
    )
    r = history_grid_result.next_row
    row_write_elapsed = history_grid_result.row_write_elapsed
    row_fill_elapsed = history_grid_result.row_fill_elapsed
    valuation_row_source_values = history_grid_result.valuation_row_source_values
    _display_m_source_map_local = history_grid_result._display_m_source_map_local
    _margin = history_grid_result._margin
    _ttm_map = history_grid_result._ttm_map
    adj_ebit_ttm_map = history_grid_result.adj_ebit_ttm_map
    adj_ebitda_map = history_grid_result.adj_ebitda_map
    adj_ebitda_ttm_map = history_grid_result.adj_ebitda_ttm_map
    adj_eps_ttm_map = history_grid_result.adj_eps_ttm_map
    adj_fcf_ttm_map = history_grid_result.adj_fcf_ttm_map
    ar_map = history_grid_result.ar_map
    assets_map = history_grid_result.assets_map
    buyback_avg_price_doc_map = history_grid_result.buyback_avg_price_doc_map
    buyback_cash_facts_map = history_grid_result.buyback_cash_facts_map
    buyback_doc_note_map = history_grid_result.buyback_doc_note_map
    buyback_map = history_grid_result.buyback_map
    buyback_shares_map = history_grid_result.buyback_shares_map
    buyback_shares_text_map = history_grid_result.buyback_shares_text_map
    buyback_ttm_map = history_grid_result.buyback_ttm_map
    bv_share_map = history_grid_result.bv_share_map
    capex_map = history_grid_result.capex_map
    capex_ttm_map = history_grid_result.capex_ttm_map
    capex_ttm_pct_source_map = history_grid_result.capex_ttm_pct_source_map
    capital_return_resolved = history_grid_result.capital_return_resolved
    cash_map = history_grid_result.cash_map
    cfo_map = history_grid_result.cfo_map
    company_operating_margin_source_map = history_grid_result.company_operating_margin_source_map
    cov_cash_display_map = history_grid_result.cov_cash_display_map
    cov_cash_map = history_grid_result.cov_cash_map
    cov_pnl_display_map = history_grid_result.cov_pnl_display_map
    cov_pnl_map = history_grid_result.cov_pnl_map
    debt_core_map = history_grid_result.debt_core_map
    debt_current_map = history_grid_result.debt_current_map
    dividend_cash_facts_map = history_grid_result.dividend_cash_facts_map
    dividend_doc_note_map = history_grid_result.dividend_doc_note_map
    dividend_map = history_grid_result.dividend_map
    dividend_ps_doc_map = history_grid_result.dividend_ps_doc_map
    dividend_ttm_map = history_grid_result.dividend_ttm_map
    ebit_map = history_grid_result.ebit_map
    ebit_margin_ttm_source_map = history_grid_result.ebit_margin_ttm_source_map
    ebitda_map = history_grid_result.ebitda_map
    ebitda_margin_ttm_source_map = history_grid_result.ebitda_margin_ttm_source_map
    ebitda_ttm_map = history_grid_result.ebitda_ttm_map
    fcf_conv_map = history_grid_result.fcf_conv_map
    fcf_margin_ttm_source_map = history_grid_result.fcf_margin_ttm_source_map
    fcf_per_share_ttm = history_grid_result.fcf_per_share_ttm
    fcf_ttm_map = history_grid_result.fcf_ttm_map
    goodwill_map = history_grid_result.goodwill_map
    gross_profit_map = history_grid_result.gross_profit_map
    history_bv_share_source_map = history_grid_result.history_bv_share_source_map
    history_capex_pct_source_map = history_grid_result.history_capex_pct_source_map
    history_current_ratio_source_map = history_grid_result.history_current_ratio_source_map
    history_debt_core_source_map = history_grid_result.history_debt_core_source_map
    history_ebit_margin_source_map = history_grid_result.history_ebit_margin_source_map
    history_ebitda_margin_source_map = history_grid_result.history_ebitda_margin_source_map
    history_eps_gaap_source_map = history_grid_result.history_eps_gaap_source_map
    history_fcf_margin_source_map = history_grid_result.history_fcf_margin_source_map
    history_fcf_per_share_ttm_source_map = history_grid_result.history_fcf_per_share_ttm_source_map
    history_fcf_source_map = history_grid_result.history_fcf_source_map
    history_fcf_ttm_source_map = history_grid_result.history_fcf_ttm_source_map
    history_gross_margin_source_map = history_grid_result.history_gross_margin_source_map
    history_net_debt_source_map = history_grid_result.history_net_debt_source_map
    history_net_income_margin_source_map = history_grid_result.history_net_income_margin_source_map
    history_owner_earnings_source_map = history_grid_result.history_owner_earnings_source_map
    int_paid_ttm_map = history_grid_result.int_paid_ttm_map
    inventory_map = history_grid_result.inventory_map
    last4_quarters_map = history_grid_result.last4_quarters_map
    liquidity_map = history_grid_result.liquidity_map
    net_debt_map = history_grid_result.net_debt_map
    net_income_label = history_grid_result.net_income_label
    net_income_map = history_grid_result.net_income_map
    net_income_margin_ttm_source_map = history_grid_result.net_income_margin_ttm_source_map
    net_income_ttm_map = history_grid_result.net_income_ttm_map
    net_lev_adj_display_map = history_grid_result.net_lev_adj_display_map
    net_lev_adj_map = history_grid_result.net_lev_adj_map
    net_lev_display_map = history_grid_result.net_lev_display_map
    net_lev_map = history_grid_result.net_lev_map
    owner_maint_capex_ratio_default = history_grid_result.owner_maint_capex_ratio_default
    pension_map = history_grid_result.pension_map
    rev_map = history_grid_result.rev_map
    rev_ttm_map = history_grid_result.rev_ttm_map
    row_operating_margin_pct = history_grid_result.row_operating_margin_pct
    row_operating_margin_ttm_pct = history_grid_result.row_operating_margin_ttm_pct
    shares_for_value_map = history_grid_result.shares_for_value_map
    shares_map = history_grid_result.shares_map
    shares_out_map = history_grid_result.shares_out_map
    tbv_share_map = history_grid_result.tbv_share_map
    total_debt_map = history_grid_result.total_debt_map
    total_equity_map = history_grid_result.total_equity_map
    valuation_price_input_available = history_grid_result.valuation_price_input_available
    valuation_render_started = history_grid_result.valuation_render_started


    debt_detail_result = render_valuation_debt_detail(
        ValuationDebtDetailRenderDeps(
            runtime={**runtime_globals, **locals(), "context_globals": globals()}
        )
    )
    r = debt_detail_result.next_row
    row_debt_detail_hdr = debt_detail_result.row_debt_detail_hdr
    tieout_diff_m = debt_detail_result.tieout_diff_m
    debt_tieout_guardrail_triggered = debt_detail_result.debt_tieout_guardrail_triggered
    latest_debt_review = debt_detail_result.latest_debt_review
    principal_total_m = debt_detail_result.principal_total_m
    carrying_total_m = debt_detail_result.carrying_total_m
    debt_current_latest_m = debt_detail_result.debt_current_latest_m
    debt_long_term_latest_m = debt_detail_result.debt_long_term_latest_m
    carrying_minus_principal_m = debt_detail_result.carrying_minus_principal_m
    near_term_m = debt_detail_result.near_term_m

    formula_core_result = render_valuation_formula_core(
        ValuationFormulaCoreRenderDeps(
            runtime={
                **runtime_globals,
                **locals(),
                "context_globals": context_globals,
                "Alignment": Alignment,
                "Border": Border,
                "CellIsRule": CellIsRule,
                "DefinedName": DefinedName,
                "Font": Font,
                "Path": Path,
                "PatternFill": PatternFill,
                "Side": Side,
                "_collapse_repeated_leading_ngram_local": _collapse_repeated_leading_ngram_local,
                "_dedupe_canonical_text_parts_local": _dedupe_canonical_text_parts_local,
                "_htmlish_to_text": _htmlish_to_text,
                "_quarter_notes_view": _quarter_notes_view,
                "_resolve_col": _resolve_col,
                "_safe_text_value": _safe_text_value,
                "_set_cell_comment_local": _set_cell_comment_local,
                "_source_backed_debt_tranches_from_slides": _source_backed_debt_tranches_from_slides,
                "adj_ebitda_map": adj_ebitda_map,
                "adj_ebitda_ttm_map": adj_ebitda_ttm_map,
                "adj_eps_ttm_map": adj_eps_ttm_map,
                "adj_fcf_ttm_map": adj_fcf_ttm_map,
                "bold": bold,
                "bv_share_map": bv_share_map,
                "capex_map": capex_map,
                "capex_ttm_map": capex_ttm_map,
                "cash_map": cash_map,
                "cfo_map": cfo_map,
                "company_overview": company_overview,
                "copy": copy,
                "debt_core_map": debt_core_map,
                "debt_tranches_latest": debt_tranches_latest,
                "ebitda_map": ebitda_map,
                "ebitda_ttm_map": ebitda_ttm_map,
                "font_size": font_size,
                "get_column_letter": get_column_letter,
                "glx_normalize_text": glx_normalize_text,
                "header_fill": header_fill,
                "input_fill": input_fill,
                "int_paid_ttm_map": int_paid_ttm_map,
                "is_anf_profile": is_anf_profile,
                "last4_quarters_map": last4_quarters_map,
                "last_col_letter": last_col_letter,
                "net_income_map": net_income_map,
                "net_lev_map": net_lev_map,
                "owner_maint_capex_ratio_default": owner_maint_capex_ratio_default,
                "pd": pd,
                "price": price,
                "qn_compact_snippet": qn_compact_snippet,
                "qs": qs,
                "quarter_columns": quarter_columns,
                "quarter_notes": quarter_notes,
                "re": re,
                "rev_map": rev_map,
                "rev_ttm_map": rev_ttm_map,
                "row_operating_margin_pct": row_operating_margin_pct,
                "row_operating_margin_ttm_pct": row_operating_margin_ttm_pct,
                "section_fill": section_fill,
                "shares_for_value_map": shares_for_value_map,
                "shares_map": shares_map,
                "shares_out_map": shares_out_map,
                "slides_debt": slides_debt,
                "tbv_share_map": tbv_share_map,
                "thin_border": thin_border,
                "ticker": ticker,
                "tieout_diff_m": tieout_diff_m,
                "valuation_grid_df": valuation_grid_df,
                "valuation_price_input_available": valuation_price_input_available,
                "wb": wb,
                "ws": ws,
            }
        )
    )
    valuation_header_row = formula_core_result.valuation_header_row
    valuation_inputs_row = formula_core_result.valuation_inputs_row
    input_label_col = formula_core_result.input_label_col
    input_value_col = formula_core_result.input_value_col
    input_basis_col = formula_core_result.input_basis_col
    input_hint_col = formula_core_result.input_hint_col
    output_label_col = formula_core_result.output_label_col
    output_value_col = formula_core_result.output_value_col
    output_interp_col = formula_core_result.output_interp_col
    market_label_col = formula_core_result.market_label_col
    market_value_col = formula_core_result.market_value_col
    market_interp_col = formula_core_result.market_interp_col
    scn_label_col = formula_core_result.scn_label_col
    scn_value_col = formula_core_result.scn_value_col
    scn_interp_col = formula_core_result.scn_interp_col
    driver_label_col = formula_core_result.driver_label_col
    driver_value_col = formula_core_result.driver_value_col
    toggle_label_col = formula_core_result.toggle_label_col
    toggle_value_col = formula_core_result.toggle_value_col
    qadj_label_col = formula_core_result.qadj_label_col
    qadj_value_col = formula_core_result.qadj_value_col
    qadj_text_col = formula_core_result.qadj_text_col
    dcf_label_col = formula_core_result.dcf_label_col
    dcf_value_col = formula_core_result.dcf_value_col
    dcf_interp_col = formula_core_result.dcf_interp_col
    grid_start = formula_core_result.grid_start
    grid_layout_width = formula_core_result.grid_layout_width
    right_stack_anchor = formula_core_result.right_stack_anchor
    date_ref = formula_core_result.date_ref
    row_price = formula_core_result.row_price
    row_asof = formula_core_result.row_asof
    row_shares_out = formula_core_result.row_shares_out
    row_shares_dil = formula_core_result.row_shares_dil
    row_net_debt = formula_core_result.row_net_debt
    row_ebitda_ttm = formula_core_result.row_ebitda_ttm
    row_adj_ebitda_ttm = formula_core_result.row_adj_ebitda_ttm
    row_fcf_ttm = formula_core_result.row_fcf_ttm
    row_adj_fcf_ttm = formula_core_result.row_adj_fcf_ttm
    row_rev_ttm = formula_core_result.row_rev_ttm
    row_eps_ttm = formula_core_result.row_eps_ttm
    row_adj_eps_ttm = formula_core_result.row_adj_eps_ttm
    row_bv = formula_core_result.row_bv
    row_tbv = formula_core_result.row_tbv
    row_tgt_ev_adj = formula_core_result.row_tgt_ev_adj
    row_tgt_ev = formula_core_result.row_tgt_ev
    row_tgt_fcf = formula_core_result.row_tgt_fcf
    row_capex_ttm = formula_core_result.row_capex_ttm
    row_int_paid_ttm = formula_core_result.row_int_paid_ttm
    row_owner_maint_ratio = formula_core_result.row_owner_maint_ratio
    row_owner_recurring = formula_core_result.row_owner_recurring
    row_owner_wc_norm = formula_core_result.row_owner_wc_norm
    row_share_mode = formula_core_result.row_share_mode
    row_out_hdr = formula_core_result.row_out_hdr
    row_mktcap = formula_core_result.row_mktcap
    row_ev = formula_core_result.row_ev
    row_implied_ev_adj = formula_core_result.row_implied_ev_adj
    row_implied_ev = formula_core_result.row_implied_ev
    row_fcff_proxy_ttm = formula_core_result.row_fcff_proxy_ttm
    row_implied_fcff = formula_core_result.row_implied_fcff
    row_equity_fcf = formula_core_result.row_equity_fcf
    row_owner_fcf_ttm = formula_core_result.row_owner_fcf_ttm
    row_owner_fcf_yield = formula_core_result.row_owner_fcf_yield
    row_eq_adj = formula_core_result.row_eq_adj
    row_eq_ev = formula_core_result.row_eq_ev
    row_eq_fcf = formula_core_result.row_eq_fcf
    row_pe = formula_core_result.row_pe
    row_pe_adj = formula_core_result.row_pe_adj
    row_ev_sales = formula_core_result.row_ev_sales
    row_pb = formula_core_result.row_pb
    row_ptbv = formula_core_result.row_ptbv
    row_mi_hdr = formula_core_result.row_mi_hdr
    row_mi_market_ev = formula_core_result.row_mi_market_ev
    row_mi_dcf_ev = formula_core_result.row_mi_dcf_ev
    row_mi_curr_wacc = formula_core_result.row_mi_curr_wacc
    row_mi_curr_gt = formula_core_result.row_mi_curr_gt
    row_mi_tbl_hdr = formula_core_result.row_mi_tbl_hdr
    row_mi_wacc_start = formula_core_result.row_mi_wacc_start
    row_mi_wacc_end = formula_core_result.row_mi_wacc_end
    row_mi_toggle = formula_core_result.row_mi_toggle
    row_dcf_hdr = formula_core_result.row_dcf_hdr
    row_dcf_start = formula_core_result.row_dcf_start
    row_dcf_g = formula_core_result.row_dcf_g
    row_dcf_gt = formula_core_result.row_dcf_gt
    row_dcf_wacc = formula_core_result.row_dcf_wacc
    row_dcf_ev = formula_core_result.row_dcf_ev
    row_dcf_eq = formula_core_result.row_dcf_eq
    row_dcf_sens_hdr = formula_core_result.row_dcf_sens_hdr
    row_dcf_sens_last_row = formula_core_result.row_dcf_sens_last_row
    row_scn_hdr = formula_core_result.row_scn_hdr
    row_scn_profile = formula_core_result.row_scn_profile
    row_scn_growth = formula_core_result.row_scn_growth
    row_scn_margin = formula_core_result.row_scn_margin
    row_scn_refi = formula_core_result.row_scn_refi
    row_scn_buyback = formula_core_result.row_scn_buyback
    row_scn_adj_ebitda = formula_core_result.row_scn_adj_ebitda
    row_scn_owner_fcf = formula_core_result.row_scn_owner_fcf
    row_scn_eq_ev = formula_core_result.row_scn_eq_ev
    row_scn_eq_fcf = formula_core_result.row_scn_eq_fcf
    row_market_hdr = formula_core_result.row_market_hdr
    row_req_adj_ebitda = formula_core_result.row_req_adj_ebitda
    row_req_adj_delta = formula_core_result.row_req_adj_delta
    row_req_fcff = formula_core_result.row_req_fcff
    row_req_fcff_delta = formula_core_result.row_req_fcff_delta
    row_req_owner_fcf = formula_core_result.row_req_owner_fcf
    row_req_owner_delta = formula_core_result.row_req_owner_delta
    row_qa = formula_core_result.row_qa
    row_drv_hdr = formula_core_result.row_drv_hdr
    row_drv_rev = formula_core_result.row_drv_rev
    row_drv_margin = formula_core_result.row_drv_margin
    row_drv_fcf = formula_core_result.row_drv_fcf
    row_drv_lev = formula_core_result.row_drv_lev
    row_toggle_hdr = formula_core_result.row_toggle_hdr
    row_toggle_reg = formula_core_result.row_toggle_reg
    row_toggle_gap = formula_core_result.row_toggle_gap
    row_toggle_lev = formula_core_result.row_toggle_lev
    row_toggle_conc = formula_core_result.row_toggle_conc
    row_qadj_hdr = formula_core_result.row_qadj_hdr
    row_qadj_ev_adj = formula_core_result.row_qadj_ev_adj
    row_qadj_ev = formula_core_result.row_qadj_ev
    row_qadj_yield = formula_core_result.row_qadj_yield
    row_dcf_end = formula_core_result.row_dcf_end
    row_hv_hdr = formula_core_result.row_hv_hdr
    row_hv_total = formula_core_result.row_hv_total
    row_hv_prof = formula_core_result.row_hv_prof
    row_hv_cash = formula_core_result.row_hv_cash
    row_hv_delev = formula_core_result.row_hv_delev
    row_hv_quality = formula_core_result.row_hv_quality
    row_hv_narr = formula_core_result.row_hv_narr
    row_hv_b1 = formula_core_result.row_hv_b1
    row_hv_b2 = formula_core_result.row_hv_b2
    row_hv_b3 = formula_core_result.row_hv_b3
    row_hv_b4 = formula_core_result.row_hv_b4
    row_hv_b5 = formula_core_result.row_hv_b5
    row_convert_hdr = formula_core_result.row_convert_hdr
    convert_header_end_col = formula_core_result.convert_header_end_col
    qa_msgs = formula_core_result.qa_msgs
    tieout_diff_m = formula_core_result.tieout_diff_m
    fair_denom = formula_core_result.fair_denom
    _normalize_thesis_bridge_basis = formula_core_result.normalize_thesis_bridge_basis
    _set_formula_name = formula_core_result.set_formula_name
    _set_cell_comment = formula_core_result.set_cell_comment
    # Hidden Value Panel values are computed here and rendered later
    # (under Notes and above sensitivity grid).
    def _pct(v: Optional[float]) -> str:
        if v is None or pd.isna(v):
            return "n/a"
        return f"{float(v) * 100:.1f}%"

    def _money_m(v: Optional[float]) -> str:
        if v is None or pd.isna(v):
            return "n/a"
        return f"${float(v) / 1e6:,.1f}m"

    def _delta_m(v: Optional[float]) -> str:
        if v is None or pd.isna(v):
            return "n/a"
        sgn = "+" if float(v) >= 0 else "-"
        return f"{sgn}${abs(float(v)) / 1e6:,.1f}m"
    hidden_value_state_runtime = {
        **runtime_globals,
        **locals(),
        "_anf_buyback_execution_is_year_or_ttm": _anf_buyback_execution_is_year_or_ttm,
        "_anf_format_year_ttm_buyback_summary": _anf_format_year_ttm_buyback_summary,
        "_anf_prior_year_quarter": _anf_prior_year_quarter,
        "_build_hidden_value_flags_fallback": _build_hidden_value_flags_fallback,
        "_delta_m": _delta_m,
        "_ensure_terminal_period": _ensure_terminal_period,
        "_extract_latest_buyback_remaining_from_sec": _extract_latest_buyback_remaining_from_sec,
        "_extract_valuation_filing_doc_text": _extract_valuation_filing_doc_text,
        "_money_m": _money_m,
        "_pct": _pct,
        "_prev_quarter_end_from_qend": _prev_quarter_end_from_qend,
        "_quarter_notes_view": _quarter_notes_view,
        "_record_writer_substage": _record_writer_substage,
        "_resolve_col": _resolve_col,
        "_sec_cache_docs_for_token_local": _sec_cache_docs_for_token_local,
        "_ttm_map": _ttm_map,
        "adj_ebitda_ttm_map": adj_ebitda_ttm_map,
        "adj_metrics": adj_metrics,
        "all_qs_ts": all_qs_ts,
        "build_hidden_value_flags": build_hidden_value_flags,
        "buyback_avg_price_doc_map": buyback_avg_price_doc_map,
        "buyback_cash_facts_map": buyback_cash_facts_map,
        "buyback_doc_note_map": buyback_doc_note_map,
        "buyback_map": buyback_map,
        "buyback_shares_map": buyback_shares_map,
        "buyback_shares_text_map": buyback_shares_text_map,
        "buyback_ttm_map": buyback_ttm_map,
        "cache_root": cache_root,
        "capital_return_resolved": capital_return_resolved,
        "cov_cash_map": cov_cash_map,
        "cov_pnl_map": cov_pnl_map,
        "date_ref": date_ref,
        "debt_credit_notes": debt_credit_notes,
        "debt_tranches": debt_tranches,
        "dividend_cash_facts_map": dividend_cash_facts_map,
        "dividend_doc_note_map": dividend_doc_note_map,
        "dividend_map": dividend_map,
        "dividend_ps_doc_map": dividend_ps_doc_map,
        "dividend_ttm_map": dividend_ttm_map,
        "ebitda_ttm_map": ebitda_ttm_map,
        "fcf_per_share_ttm": fcf_per_share_ttm,
        "fcf_ttm_map": fcf_ttm_map,
        "flags_audit_df": flags_audit_df,
        "flags_df": flags_df,
        "glx_normalize_text": glx_normalize_text,
        "hist": hist,
        "is_anf_profile": is_anf_profile,
        "last4_quarters_map": last4_quarters_map,
        "leverage_df": leverage_df,
        "manifest_df": manifest_df,
        "net_debt_map": net_debt_map,
        "price": price,
        "promises": promises,
        "qs": qs,
        "quarter_notes": quarter_notes,
        "rev_ttm_map": rev_ttm_map,
        "shares_for_value_map": shares_for_value_map,
        "shares_out_map": shares_out_map,
        "signals_base_df": signals_base_df,
        "strip_html": strip_html,
        "ticker": ticker,
    }
    hidden_value_state = build_valuation_hidden_value_state(
        ValuationHiddenValueStateDeps(runtime=hidden_value_state_runtime)
    )
    hv_scores = hidden_value_state.hv_scores
    hv_obs = hidden_value_state.hv_obs
    hv_buybacks = hidden_value_state.hv_buybacks
    hv_buybacks_note = hidden_value_state.hv_buybacks_note
    hv_dividends = hidden_value_state.hv_dividends
    hv_dividends_note = hidden_value_state.hv_dividends_note

    hv_panel_label_col = 14  # N; consumed by final cross-panel layout.
    hv_panel_val_col = 18  # R; consumed by final cross-panel layout.
    hidden_value_render_runtime = {
        **runtime_globals,
        **locals(),
        "context_globals": context_globals,
        "_anf_visible_quarter_label": _anf_visible_quarter_label,
        "_build_hidden_value_flags_fallback": _build_hidden_value_flags_fallback,
        "_estimate_wrapped_row_height": _estimate_wrapped_row_height,
        "_fmt_short_money_value_local": _fmt_short_money_value_local,
        "adj_ebit_ttm_map": adj_ebit_ttm_map,
        "adj_ebitda_ttm_map": adj_ebitda_ttm_map,
        "adj_metrics": adj_metrics,
        "bold": bold,
        "build_hidden_value_flags": build_hidden_value_flags,
        "buyback_map": buyback_map,
        "buyback_ttm_map": buyback_ttm_map,
        "cov_cash_display_map": cov_cash_display_map,
        "cov_pnl_display_map": cov_pnl_display_map,
        "cov_pnl_map": cov_pnl_map,
        "ctx_ref": ctx_ref,
        "debt_tranches": debt_tranches,
        "dividend_ttm_map": dividend_ttm_map,
        "flags_audit_df": flags_audit_df,
        "flags_df": flags_df,
        "glx_normalize_text": glx_normalize_text,
        "header_fill": header_fill,
        "hist": hist,
        "hv_buybacks": hv_buybacks,
        "hv_buybacks_note": hv_buybacks_note,
        "hv_dividends": hv_dividends,
        "hv_dividends_note": hv_dividends_note,
        "hv_obs": hv_obs,
        "hv_scores": hv_scores,
        "is_anf_profile": is_anf_profile,
        "leverage_df": leverage_df,
        "net_lev_adj_display_map": net_lev_adj_display_map,
        "net_lev_adj_map": net_lev_adj_map,
        "net_lev_display_map": net_lev_display_map,
        "price": price,
        "qs_ts": qs_ts,
        "section_fill": section_fill,
        "signals_base_df": signals_base_df,
        "thin_border": thin_border,
        "valuation_header_row": valuation_header_row,
        "ws": ws,
    }
    hidden_value_render_result = render_valuation_hidden_value_panel(
        ValuationHiddenValueRenderDeps(runtime=hidden_value_render_runtime)
    )
    row_hv_hdr_dyn = hidden_value_render_result.row_score_panel_header
    row_hv_obs_hdr_dyn = hidden_value_render_result.row_operating_signals_header
    row_hv_cap_hdr_dyn = hidden_value_render_result.row_capital_return_header
    row_hv_dividends_note_dyn = hidden_value_render_result.row_dividends_note
    visible_hv_flags_hdr_row = hidden_value_render_result.row_flags_header

    trend_flags_render_result = render_valuation_trend_flags_panel(
        ValuationTrendFlagsRenderDeps(
            runtime={
                **runtime_globals,
                **locals(),
                "context_globals": context_globals,
                "font_size": font_size,
                "header_size": header_size,
                "_record_writer_elapsed": _record_writer_elapsed,
                "_record_writer_substage": _record_writer_substage,
                "_updated_font": _updated_font,
            }
        )
    )
    row_trend_hdr = trend_flags_render_result.row_trend_hdr
    row_flags_hdr = trend_flags_render_result.row_flags_hdr
    panel_col = trend_flags_render_result.panel_col
    panel_row = trend_flags_render_result.next_panel_row

    valuation_guidance_render_result = render_valuation_guidance_panel(
        ValuationGuidanceRenderDeps(
            runtime={
                **runtime_globals,
                **locals(),
                "context_globals": context_globals,
                "_audit_view": _audit_view,
                "_ensure_terminal_period": _ensure_terminal_period,
                "_extract_45z_monetization_target_display": _extract_45z_monetization_target_display,
                "_extract_money_targets_for_display": _extract_money_targets_for_display,
                "_extract_pbi_target_display": _extract_pbi_target_display,
                "_first_existing_material_dir": _first_existing_material_dir,
                "_fmt_short_money_value_local": _fmt_short_money_value_local,
                "_gpre_commercial_setup_records_shared": _gpre_commercial_setup_records_shared,
                "_gpre_local_bmo_conference_path_shared": _gpre_local_bmo_conference_path_shared,
                "_gpre_local_bmo_conference_text_shared": _gpre_local_bmo_conference_text_shared,
                "_gpre_local_bofa_conference_path_shared": _gpre_local_bofa_conference_path_shared,
                "_gpre_local_bofa_conference_text_shared": _gpre_local_bofa_conference_text_shared,
                "_gpre_local_stephens_conference_path_shared": _gpre_local_stephens_conference_path_shared,
                "_gpre_local_stephens_conference_raw_path_shared": _gpre_local_stephens_conference_raw_path_shared,
                "_gpre_local_stephens_conference_raw_text_shared": _gpre_local_stephens_conference_raw_text_shared,
                "_gpre_local_stephens_conference_text_shared": _gpre_local_stephens_conference_text_shared,
                "_gpre_normalize_metric_label": _gpre_normalize_metric_label,
                "_load_profile_slide_signals": _load_profile_slide_signals,
                "_pbi_guidance_period_label_from_text": _pbi_guidance_period_label_from_text,
                "_pbi_repair_guidance_period_meta": _pbi_repair_guidance_period_meta,
                "_pbi_structured_strategy_items_for_qd": _pbi_structured_strategy_items_for_qd,
                "_period_label_to_norm": _period_label_to_norm,
                "_profile_slide_signals_for_quarter": _profile_slide_signals_for_quarter,
                "_promises_view": _promises_view,
                "_quarter_notes_view": _quarter_notes_view,
                "_read_cached_doc_text": _read_cached_doc_text,
                "_read_local_doc_text_shared": _read_local_doc_text_shared,
                "_resolve_cached_doc_path": _resolve_cached_doc_path,
                "_resolve_col": _resolve_col,
                "_sec_docs_for_accession": _sec_docs_for_accession,
                "_set_cell_comment_local": _set_cell_comment_local,
                "_slide_signal_noise": _slide_signal_noise,
                "_submission_recent_row_quarter": _submission_recent_row_quarter,
                "_submission_recent_rows": _submission_recent_rows,
                "audit": audit,
                "cache_root": cache_root,
                "hist": hist,
                "is_gpre_profile": is_gpre_profile,
                "is_pbi_profile": is_pbi_profile,
                "promise_progress": promise_progress,
                "promises": promises,
                "quarter_notes": quarter_notes,
                "slides_guidance": slides_guidance,
                "ui_state": ui_state,
            }
        )
    )
    panel_col_start = valuation_guidance_render_result.panel_col_start
    panel_col_end = valuation_guidance_render_result.panel_col_end
    additive_panel_end = valuation_guidance_render_result.additive_panel_end
    panel_row_start = valuation_guidance_render_result.panel_row_start
    col_metric_start = valuation_guidance_render_result.col_metric_start
    col_stated_start = valuation_guidance_render_result.col_stated_start
    col_horizon_start = valuation_guidance_render_result.col_horizon_start
    col_guidance_start = valuation_guidance_render_result.col_guidance_start
    col_exact_start = valuation_guidance_render_result.col_exact_start
    side_panel_style = valuation_guidance_render_result.side_panel_style
    guidance_snapshot_header_rows = valuation_guidance_render_result.guidance_snapshot_header_rows
    _overlaps = valuation_guidance_render_result.overlaps
    row_ptr = valuation_guidance_render_result.row_ptr
    valuation_operating_thesis_render_result = render_valuation_operating_thesis_panels(
        ValuationOperatingThesisRenderDeps(
            runtime={
                **runtime_globals,
                **locals(),
                "context_globals": context_globals,
                "_build_operating_driver_rows": _build_operating_driver_rows,
                "_resolve_thesis_fy_base": _resolve_thesis_fy_base,
                "_set_cell_comment_local": _set_cell_comment_local,
            }
        )
    )
    row_operating_hdr = valuation_operating_thesis_render_result.row_operating_hdr
    row_operating_end = valuation_operating_thesis_render_result.row_operating_end
    row_thesis_hdr = valuation_operating_thesis_render_result.row_thesis_hdr
    row_thesis_end = valuation_operating_thesis_render_result.row_thesis_end

    sensitivity_heatmap_render_result = render_valuation_sensitivity_heatmaps(
        ValuationSensitivityHeatmapRenderDeps(
            runtime={
                **runtime_globals,
                **locals(),
                "context_globals": context_globals,
            }
        )
    )
    grid_start = sensitivity_heatmap_render_result.grid_start
    grid_col_start = sensitivity_heatmap_render_result.grid_col_start
    grid_layout_width = sensitivity_heatmap_render_result.grid_layout_width

    apply_valuation_final_layout(
        ValuationFinalLayoutDeps(
            runtime={
                **runtime_globals,
                **locals(),
                "context_globals": context_globals,
            }
        )
    )
