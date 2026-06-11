"""Worksheet render adapter for Valuation sensitivity grid and heatmap fills."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, MutableMapping, Optional, Tuple

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ValuationSensitivityHeatmapRenderDeps:
    runtime: MutableMapping[str, Any]


@dataclass(frozen=True)
class ValuationSensitivityHeatmapRenderResult:
    grid_start: int
    grid_col_start: int
    grid_layout_width: int
    grid_last_row: Optional[int]
    grid_last_col: Optional[int]
    heatmap_rows_touched: int
    source_override_rows_touched: int


def render_valuation_sensitivity_heatmaps(
    deps: ValuationSensitivityHeatmapRenderDeps,
) -> ValuationSensitivityHeatmapRenderResult:
    __rt = deps.runtime
    context_globals = dict(__rt.get("context_globals") or {})

    def _rt_get(name: str) -> Any:
        if name in __rt:
            return __rt[name]
        if name in context_globals:
            return context_globals[name]
        return globals().get(name)

    _display_m_source_map_local = _rt_get("_display_m_source_map_local")
    _hidden_source_comparison_metric = _rt_get("_hidden_source_comparison_metric")
    _quarterly_color_basis_for_label = _rt_get("_quarterly_color_basis_for_label")
    _quarterly_color_metric_from_series = _rt_get("_quarterly_color_metric_from_series")
    _quarterly_row_color_policy = _rt_get("_quarterly_row_color_policy")
    capex_ttm_pct_source_map = _rt_get("capex_ttm_pct_source_map")
    company_operating_margin_source_map = _rt_get("company_operating_margin_source_map")
    cov_cash_display_map = _rt_get("cov_cash_display_map")
    cov_pnl_display_map = _rt_get("cov_pnl_display_map")
    ebit_margin_ttm_source_map = _rt_get("ebit_margin_ttm_source_map")
    ebitda_margin_ttm_source_map = _rt_get("ebitda_margin_ttm_source_map")
    fcf_conv_map = _rt_get("fcf_conv_map")
    fcf_margin_ttm_source_map = _rt_get("fcf_margin_ttm_source_map")
    grid_layout_width = int(_rt_get("grid_layout_width") or 4)
    header_fill = _rt_get("header_fill")
    history_bv_share_source_map = _rt_get("history_bv_share_source_map")
    history_capex_pct_source_map = _rt_get("history_capex_pct_source_map")
    history_current_ratio_source_map = _rt_get("history_current_ratio_source_map")
    history_debt_core_source_map = _rt_get("history_debt_core_source_map")
    history_ebit_margin_source_map = _rt_get("history_ebit_margin_source_map")
    history_ebitda_margin_source_map = _rt_get("history_ebitda_margin_source_map")
    history_eps_gaap_source_map = _rt_get("history_eps_gaap_source_map")
    history_fcf_margin_source_map = _rt_get("history_fcf_margin_source_map")
    history_fcf_per_share_ttm_source_map = _rt_get("history_fcf_per_share_ttm_source_map")
    history_fcf_source_map = _rt_get("history_fcf_source_map")
    history_fcf_ttm_source_map = _rt_get("history_fcf_ttm_source_map")
    history_gross_margin_source_map = _rt_get("history_gross_margin_source_map")
    history_net_debt_source_map = _rt_get("history_net_debt_source_map")
    history_net_income_margin_source_map = _rt_get("history_net_income_margin_source_map")
    history_owner_earnings_source_map = _rt_get("history_owner_earnings_source_map")
    last_col = _rt_get("last_col")
    net_income_label = _rt_get("net_income_label")
    net_income_margin_ttm_source_map = _rt_get("net_income_margin_ttm_source_map")
    net_lev_adj_map = _rt_get("net_lev_adj_map")
    net_lev_map = _rt_get("net_lev_map")
    pd = _rt_get("pd")
    qs_ts = _rt_get("qs_ts")
    row_operating_margin_pct = _rt_get("row_operating_margin_pct")
    section_fill = _rt_get("section_fill")
    start_col = _rt_get("start_col")
    valuation_grid_df = _rt_get("valuation_grid_df")
    valuation_hidden_comparison_metric = _rt_get("valuation_hidden_comparison_metric")
    valuation_row_source_values = _rt_get("valuation_row_source_values")
    ws = _rt_get("ws")
    bold = _rt_get("bold")

    # Keep the sensitivity grid one row below the denominator row so the
    # per-share mode hint remains readable without changing the surrounding layout.
    grid_start = 217
    grid_col_start = 2  # B
    grid_col_letter = get_column_letter(grid_col_start)
    grid_last_row: Optional[int] = None
    grid_last_col: Optional[int] = None
    if valuation_grid_df is not None and not valuation_grid_df.empty:
        grid_headers = list(valuation_grid_df.columns)[:4]
        grid_last_col = grid_col_start + max(0, len(grid_headers) - 1)
        grid_last_row = grid_start + 2 + len(valuation_grid_df)
        for mrange in list(ws.merged_cells.ranges):
            if (
                mrange.max_row >= grid_start
                and mrange.min_row <= grid_last_row
                and mrange.max_col >= grid_col_start
                and mrange.min_col <= grid_last_col
            ):
                try:
                    ws.unmerge_cells(str(mrange))
                except Exception:
                    pass
        ws[f"{grid_col_letter}{grid_start}"] = "Valuation Sensitivity Grid"
        ws[f"{grid_col_letter}{grid_start}"].font = bold
        ws[f"{grid_col_letter}{grid_start}"].fill = section_fill
        for i, hname in enumerate(grid_headers):
            cell = ws.cell(row=grid_start + 1, column=grid_col_start + i, value=hname)
            cell.font = bold
            cell.fill = header_fill
        r0 = grid_start + 2
        for ridx, (_, rowv) in enumerate(valuation_grid_df.iterrows()):
            rr = r0 + ridx
            for cidx, hname in enumerate(grid_headers):
                val = rowv.get(hname)
                ws.cell(row=rr, column=grid_col_start + cidx, value=None if pd.isna(val) else val)
                if hname.endswith("_price"):
                    ws.cell(row=rr, column=grid_col_start + cidx).number_format = "$#,##0.00"
                elif hname.endswith("multiple"):
                    ws.cell(row=rr, column=grid_col_start + cidx).number_format = "0.00x"
                else:
                    ws.cell(row=rr, column=grid_col_start + cidx).number_format = "#,##0.000"

    # Keep valuation top and historical rows visually consistent.
    ws.row_dimensions[1].height = 18.0
    for rr in range(7, 118):
        ws.row_dimensions[rr].height = 19.5

    # Keep default row heights in valuation area.

    # Conditional formatting for YoY rows (Okabe-Ito colors)
    def _apply_yoy_cf(row_idx: int) -> None:
        start = get_column_letter(start_col)
        end = get_column_letter(last_col)
        rng = f"{start}{row_idx}:{end}{row_idx}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=["-0.15"], fill=PatternFill("solid", fgColor="A63A00")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["-0.15", "-0.05"], fill=PatternFill("solid", fgColor="D55E00")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["-0.05", "0.05"], fill=PatternFill("solid", fgColor="DDDDDD")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["0.05", "0.15"], fill=PatternFill("solid", fgColor="9BD3F5")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0.15"], fill=PatternFill("solid", fgColor="2F80ED")))

    def _bucket_fill(v: Any) -> Optional[PatternFill]:
        num = pd.to_numeric(v, errors="coerce")
        if pd.isna(num):
            return None
        x = float(num)
        if x <= -0.15:
            return PatternFill("solid", fgColor="A63A00")
        if x <= -0.05:
            return PatternFill("solid", fgColor="D55E00")
        if x <= 0.05:
            return PatternFill("solid", fgColor="DDDDDD")
        if x <= 0.15:
            return PatternFill("solid", fgColor="9BD3F5")
        return PatternFill("solid", fgColor="2F80ED")

    def _coerce_pct_cell_value(v: Any) -> Any:
        if not isinstance(v, str):
            return v
        s = str(v).strip()
        if not s or "%" not in s:
            return v
        s = s.replace("\u2212", "-").replace("%", "").replace(" ", "").replace(",", ".")
        try:
            return float(s) / 100.0
        except Exception:
            return v

    def _find_row_idx_by_label(label: str) -> Optional[int]:
        want = str(label or "").strip().lower()
        if not want:
            return None
        for rr in range(1, ws.max_row + 1):
            if str(ws.cell(row=rr, column=1).value or "").strip().lower() == want:
                return rr
        return None

    def _as_float(v: Any) -> Optional[float]:
        num = pd.to_numeric(v, errors="coerce")
        if pd.isna(num):
            return None
        return float(num)

    def _valuation_hidden_comparison_metric_local(
        row_label: str,
        *,
        current_q: pd.Timestamp,
        current_value: Any,
        visible_idx: int,
        comparison_basis: str,
        directionality: str,
    ) -> Optional[float]:
        source_map = valuation_row_source_values.get(str(row_label or "")) or {}
        return valuation_hidden_comparison_metric(
            source_map,
            current_q=current_q,
            current_value=current_value,
            visible_idx=visible_idx,
            comparison_basis=comparison_basis,
            directionality=directionality,
        )

    def _row_is_percent(rr: int) -> bool:
        lbl = str(ws.cell(row=rr, column=1).value or "").strip().lower()
        if "%" in lbl:
            return True
        for cc in range(start_col, last_col + 1):
            c = ws.cell(row=rr, column=cc)
            if _as_float(c.value) is not None:
                return "%" in str(c.number_format or "")
        return False

    valuation_subheaders = {
        "Top line",
        "Margins",
        "Core operating",
        "Adjusted operating",
        "GAAP earnings",
        "Core cash flow",
        "Adjusted / derived",
        "Cash-flow quality",
        "Capital return / financing",
        "Net debt position",
        "Supplemental net cash / lease-adjusted view",
        "Coverage / leverage",
        "Revolver / liquidity",
        "Short-term liquidity",
        "Share count",
        "Per-share earnings",
        "Per-share value",
        "Market-linked",
    }

    heatmap_rows_touched = 0
    source_override_rows_touched = 0

    # Section-wide heatmap coloring uses a shared metric-aware policy.
    section_labels = ["Operating", "Cash Flow", "Leverage & Liquidity", "Equity / Per-share"]
    section_rows: List[Tuple[str, int]] = []
    for sl in section_labels:
        sr = _find_row_idx_by_label(sl)
        if sr is not None:
            section_rows.append((sl, sr))
    section_rows = sorted(section_rows, key=lambda x: x[1])
    terminal_row = _find_row_idx_by_label("Debt Detail (latest)") or ws.max_row
    for idx, (_sl, sr) in enumerate(section_rows):
        er = section_rows[idx + 1][1] - 1 if idx + 1 < len(section_rows) else terminal_row - 1
        current_subsection = ""
        for rr in range(sr + 1, er + 1):
            row_label = str(ws.cell(row=rr, column=1).value or "").strip()
            if not row_label or row_label in section_labels:
                continue
            if row_label in valuation_subheaders:
                current_subsection = row_label
                continue
            policy = _quarterly_row_color_policy(
                row_label,
                section_label=_sl,
                subsection_label=current_subsection,
            )
            row_values: List[Any] = []
            for cc in range(start_col, last_col + 1):
                c = ws.cell(row=rr, column=cc)
                c.value = _coerce_pct_cell_value(c.value)
                row_values.append(c.value)
            for idx_cc, cc in enumerate(range(start_col, last_col + 1)):
                c = ws.cell(row=rr, column=cc)
                metric = _quarterly_color_metric_from_series(
                    row_values,
                    idx_cc,
                    comparison_basis=policy.comparison_basis,
                    directionality=policy.directionality,
                )
                if metric is None and idx_cc < len(qs_ts):
                    metric = _valuation_hidden_comparison_metric_local(
                        row_label,
                        current_q=pd.Timestamp(qs_ts[idx_cc]).normalize(),
                        current_value=c.value,
                        visible_idx=idx_cc,
                        comparison_basis=policy.comparison_basis,
                        directionality=policy.directionality,
                    )
                if metric is not None:
                    bf = _bucket_fill(metric)
                    if bf is not None:
                        c.fill = bf
                        heatmap_rows_touched += 1

    # Operating margin is a derived row, and the first visible fiscal-year
    # quarters sometimes need prior-year history that lives outside the
    # visible revenue/EBIT maps. Apply the same hidden-source comparison
    # directly from History_Q as a narrow fallback; values stay unchanged.
    if row_operating_margin_pct and company_operating_margin_source_map:
        for idx_cc, cc in enumerate(range(start_col, last_col + 1)):
            c = ws.cell(row=row_operating_margin_pct, column=cc)
            metric = _hidden_source_comparison_metric(
                current_key=pd.Timestamp(qs_ts[idx_cc]).normalize(),
                current_value=c.value,
                visible_idx=idx_cc,
                comparison_basis="yoy",
                directionality="higher_better",
                source_values=company_operating_margin_source_map,
            )
            if metric is None:
                continue
            bf = _bucket_fill(metric)
            if bf is not None:
                c.fill = bf
                heatmap_rows_touched += 1

    valuation_hidden_source_fill_overrides: Dict[str, Tuple[Dict[pd.Timestamp, Any], str]] = {
        "Gross margin %": (history_gross_margin_source_map, "higher_better"),
        "Operating margin (TTM)": (ebit_margin_ttm_source_map, "higher_better"),
        "EBITDA margin %": (history_ebitda_margin_source_map, "higher_better"),
        "EBITDA margin (TTM)": (ebitda_margin_ttm_source_map, "higher_better"),
        "Adj EBITDA margin %": (history_ebitda_margin_source_map, "higher_better"),
        "Adj EBITDA margin (TTM)": (ebitda_margin_ttm_source_map, "higher_better"),
        "EBIT margin %": (history_ebit_margin_source_map, "higher_better"),
        "EBIT margin (TTM)": (ebit_margin_ttm_source_map, "higher_better"),
        f"{net_income_label} margin %": (history_net_income_margin_source_map, "higher_better"),
        f"{net_income_label} margin (TTM)": (net_income_margin_ttm_source_map, "higher_better"),
        "Capex % of revenue": (history_capex_pct_source_map, "lower_better"),
        "Capex % of revenue (TTM)": (capex_ttm_pct_source_map, "lower_better"),
        "FCF (CFO-Capex)": (_display_m_source_map_local(history_fcf_source_map), "higher_better"),
        "FCF (TTM)": (_display_m_source_map_local(history_fcf_ttm_source_map), "higher_better"),
        "Owner earnings (proxy)": (_display_m_source_map_local(history_owner_earnings_source_map), "higher_better"),
        "FCF margin %": (history_fcf_margin_source_map, "higher_better"),
        "FCF margin (TTM)": (fcf_margin_ttm_source_map, "higher_better"),
        "Debt (core borrowings)": (_display_m_source_map_local(history_debt_core_source_map), "lower_better"),
        "Net debt (core borrowings)": (_display_m_source_map_local(history_net_debt_source_map), "lower_better"),
        "Net leverage": (net_lev_map, "lower_better"),
        "Net leverage (Adj)": (net_lev_adj_map, "lower_better"),
        "Interest coverage (P&L TTM)": (cov_pnl_display_map, "higher_better"),
        "Cash interest coverage (TTM)": (cov_cash_display_map, "higher_better"),
        "FCF conversion (TTM)": (fcf_conv_map, "higher_better"),
        "Current ratio": (history_current_ratio_source_map, "higher_better"),
        "EPS (GAAP)": (history_eps_gaap_source_map, "higher_better"),
        "BV/share": (history_bv_share_source_map, "higher_better"),
        "FCF/share (TTM)": (history_fcf_per_share_ttm_source_map, "higher_better"),
    }
    for row_label, (source_map, direction) in valuation_hidden_source_fill_overrides.items():
        if not source_map:
            continue
        rr = _find_row_idx_by_label(row_label)
        if rr is None:
            continue
        basis = _quarterly_color_basis_for_label(row_label)
        for idx_cc, cc in enumerate(range(start_col, last_col + 1)):
            c = ws.cell(row=rr, column=cc)
            metric = _hidden_source_comparison_metric(
                current_key=pd.Timestamp(qs_ts[idx_cc]).normalize(),
                current_value=c.value,
                visible_idx=idx_cc,
                comparison_basis=basis,
                directionality=direction,
                source_values=source_map,
            )
            if metric is None:
                continue
            bf = _bucket_fill(metric)
            if bf is not None:
                c.fill = bf
                source_override_rows_touched += 1

    return ValuationSensitivityHeatmapRenderResult(
        grid_start=grid_start,
        grid_col_start=grid_col_start,
        grid_layout_width=grid_layout_width,
        grid_last_row=grid_last_row,
        grid_last_col=grid_last_col,
        heatmap_rows_touched=heatmap_rows_touched,
        source_override_rows_touched=source_override_rows_touched,
    )
