"""Worksheet render adapter for the Valuation Trend and Red/Green flags panels."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, MutableMapping, Optional

import pandas as pd
from openpyxl.styles import Alignment, PatternFill


@dataclass(frozen=True)
class ValuationTrendFlagsRenderDeps:
    runtime: MutableMapping[str, Any]


@dataclass(frozen=True)
class ValuationTrendFlagsRenderResult:
    row_trend_hdr: int
    row_trend_end: int
    row_flags_hdr: int
    row_flags_end: int
    panel_col: int
    panel_row_start: int
    next_panel_row: int


def render_valuation_trend_flags_panel(
    deps: ValuationTrendFlagsRenderDeps,
) -> ValuationTrendFlagsRenderResult:
    __rt = deps.runtime
    context_globals = dict(__rt.get("context_globals") or {})

    def _rt_get(name: str) -> Any:
        if name in __rt:
            return __rt[name]
        if name in context_globals:
            return context_globals[name]
        return globals().get(name)

    _anf_prior_year_quarter = _rt_get("_anf_prior_year_quarter")
    _delta_m = _rt_get("_delta_m")
    _margin = _rt_get("_margin")
    _money_m = _rt_get("_money_m")
    _net_debt_yoy_flag_label_and_status_for_position = _rt_get(
        "_net_debt_yoy_flag_label_and_status_for_position"
    )
    _pct = _rt_get("_pct")
    _ttm_map = _rt_get("_ttm_map")
    all_qs_ts = _rt_get("all_qs_ts")
    ar_map = _rt_get("ar_map")
    assets_map = _rt_get("assets_map")
    bold = _rt_get("bold")
    buyback_ttm_map = _rt_get("buyback_ttm_map")
    cfo_map = _rt_get("cfo_map")
    cov_cash_map = _rt_get("cov_cash_map")
    debt_core_map = _rt_get("debt_core_map")
    dividend_ttm_map = _rt_get("dividend_ttm_map")
    ebit_map = _rt_get("ebit_map")
    ebitda_ttm_map = _rt_get("ebitda_ttm_map")
    fcf_ttm_map = _rt_get("fcf_ttm_map")
    goodwill_map = _rt_get("goodwill_map")
    gross_profit_map = _rt_get("gross_profit_map")
    header_fill = _rt_get("header_fill")
    hidden_value_render_result = _rt_get("hidden_value_render_result")
    inventory_map = _rt_get("inventory_map")
    is_anf_profile = bool(_rt_get("is_anf_profile"))
    liquidity_map = _rt_get("liquidity_map")
    net_debt_map = _rt_get("net_debt_map")
    net_income_map = _rt_get("net_income_map")
    net_income_ttm_map = _rt_get("net_income_ttm_map")
    net_lev_map = _rt_get("net_lev_map")
    pension_map = _rt_get("pension_map")
    qs = _rt_get("qs")
    rev_map = _rt_get("rev_map")
    section_fill = _rt_get("section_fill")
    shares_for_value_map = _rt_get("shares_for_value_map")
    shares_out_map = _rt_get("shares_out_map")
    total_debt_map = _rt_get("total_debt_map")
    total_equity_map = _rt_get("total_equity_map")
    ws = _rt_get("ws")

    q0 = pd.Timestamp(qs[-1]) if qs else None
    last4_qs = [pd.Timestamp(x) for x in qs[-4:]] if qs else []
    panel_col = 1  # A
    panel_row = hidden_value_render_result.next_panel_row
    panel_row_start = panel_row
    asof_txt = str(q0.date()) if q0 is not None else "N/A"

    def _clear_merges(row_idx: int, col_start: int, col_end: int) -> None:
        for mrange in list(ws.merged_cells.ranges):
            if mrange.min_row <= row_idx <= mrange.max_row and not (mrange.max_col < col_start or mrange.min_col > col_end):
                try:
                    ws.unmerge_cells(str(mrange))
                except Exception:
                    pass

    def _status_fill(status: str) -> PatternFill:
        s = str(status or "").upper()
        if s == "PASS":
            return PatternFill("solid", fgColor="D9EAF7")
        if s == "WARN":
            return PatternFill("solid", fgColor="FFF2CC")
        if s == "FAIL":
            return PatternFill("solid", fgColor="F4CCCC")
        return PatternFill("solid", fgColor="F2F2F2")

    def _yoy_pct(src: Dict[pd.Timestamp, Any], qref: Optional[pd.Timestamp]) -> Optional[float]:
        if qref is None:
            return None
        q_norm = pd.Timestamp(qref).normalize()
        cur = src.get(q_norm)
        if is_anf_profile:
            prev_q = _anf_prior_year_quarter(q_norm, all_qs_ts)
            prev = src.get(prev_q) if prev_q is not None else None
        else:
            prev = src.get(q_norm - pd.DateOffset(years=1))
        if cur is None or prev in (None, 0):
            return None
        return (float(cur) - float(prev)) / abs(float(prev))

    def _ttm_yoy(ttm_map: Dict[pd.Timestamp, Any], qref: Optional[pd.Timestamp]) -> Optional[float]:
        if qref is None:
            return None
        q_norm = pd.Timestamp(qref).normalize()
        cur = ttm_map.get(q_norm)
        if is_anf_profile:
            prev_q = _anf_prior_year_quarter(q_norm, all_qs_ts)
            prev = ttm_map.get(prev_q) if prev_q is not None else None
        else:
            prev = ttm_map.get(q_norm - pd.DateOffset(years=1))
        if cur is None or prev in (None, 0):
            return None
        return (float(cur) - float(prev)) / abs(float(prev))

    def _fmt_pp(v: Optional[float]) -> str:
        if v is None or pd.isna(v):
            return "N/A"
        return f"{float(v):+.1f}pp"

    def _fmt_pct_delta(v: Optional[float]) -> str:
        if v is None or pd.isna(v):
            return "N/A"
        return f"{float(v) * 100:+.1f}%"

    def _trend_direction(v: Optional[float], up_thr: float, down_thr: float) -> str:
        if v is None or pd.isna(v):
            return "n/a"
        if float(v) > up_thr:
            return "up"
        if float(v) < down_thr:
            return "down"
        return "flat"

    def _margin_direction_pp(v_pp: Optional[float]) -> str:
        if v_pp is None or pd.isna(v_pp):
            return "n/a"
        if float(v_pp) > 0.5:
            return "expanding"
        if float(v_pp) < -0.5:
            return "compressing"
        return "stable"

    # Trend/Delta panel.
    row_trend_hdr = panel_row
    ws.cell(row=panel_row, column=panel_col, value="Trend/Δ (last 4Q)").font = bold
    for cc in range(panel_col, panel_col + 4):
        ws.cell(row=panel_row, column=cc).fill = section_fill
    panel_row += 1
    for i, htxt in enumerate(["Metric", "Δ", "Direction", "As-of"], start=0):
        hc = ws.cell(row=panel_row, column=panel_col + i, value=htxt)
        hc.font = bold
        hc.fill = header_fill
        hc.alignment = Alignment(horizontal="center", vertical="center")
    panel_row += 1

    rev_delta_4q = None
    gross_margin_delta_pp = None
    op_margin_delta_pp = None
    net_margin_delta_pp = None
    if len(last4_qs) == 4:
        q_start = last4_qs[0]
        q_end = last4_qs[-1]
        rev_s = rev_map.get(q_start)
        rev_e = rev_map.get(q_end)
        if rev_s not in (None, 0) and rev_e is not None:
            rev_delta_4q = (float(rev_e) - float(rev_s)) / abs(float(rev_s))
        gm_map = _margin(gross_profit_map, rev_map)
        om_map = _margin(ebit_map, rev_map)
        nm_map = _margin(net_income_map, rev_map)
        gm_s = gm_map.get(q_start)
        gm_e = gm_map.get(q_end)
        om_s = om_map.get(q_start)
        om_e = om_map.get(q_end)
        nm_s = nm_map.get(q_start)
        nm_e = nm_map.get(q_end)
        if gm_s is not None and gm_e is not None:
            gross_margin_delta_pp = (float(gm_e) - float(gm_s)) * 100.0
        if om_s is not None and om_e is not None:
            op_margin_delta_pp = (float(om_e) - float(om_s)) * 100.0
        if nm_s is not None and nm_e is not None:
            net_margin_delta_pp = (float(nm_e) - float(nm_s)) * 100.0

    fcf_ttm_y = _ttm_yoy(fcf_ttm_map, q0)
    net_debt_yoy_delta_q = None
    if q0 is not None:
        if is_anf_profile:
            prev_q0 = _anf_prior_year_quarter(q0, all_qs_ts)
            if prev_q0 is not None and net_debt_map.get(q0) is not None and net_debt_map.get(prev_q0) is not None:
                net_debt_yoy_delta_q = float(net_debt_map.get(q0)) - float(net_debt_map.get(prev_q0))
        elif net_debt_map.get(q0) is not None and net_debt_map.get(pd.Timestamp(q0) - pd.DateOffset(years=1)) is not None:
            net_debt_yoy_delta_q = float(net_debt_map.get(q0)) - float(net_debt_map.get(pd.Timestamp(q0) - pd.DateOffset(years=1)))
    shares_src_map = shares_out_map if any(v is not None for v in shares_out_map.values()) else shares_for_value_map
    shares_yoy_q = _yoy_pct(shares_src_map, q0)

    trend_rows = [
        ("Revenue", _fmt_pct_delta(rev_delta_4q), _trend_direction(rev_delta_4q, 0.01, -0.01)),
        ("Gross margin", _fmt_pp(gross_margin_delta_pp), _margin_direction_pp(gross_margin_delta_pp)),
        ("Operating margin", _fmt_pp(op_margin_delta_pp), _margin_direction_pp(op_margin_delta_pp)),
        ("Net margin", _fmt_pp(net_margin_delta_pp), _margin_direction_pp(net_margin_delta_pp)),
        ("FCF TTM (YoY)", _fmt_pct_delta(fcf_ttm_y), _trend_direction(fcf_ttm_y, 0.10, -0.10)),
        ("Net debt (YoY Δ)", _delta_m(net_debt_yoy_delta_q), _trend_direction(net_debt_yoy_delta_q, 0.0, 0.0)),
        ("Shares out (YoY)", _fmt_pct_delta(shares_yoy_q), _trend_direction(shares_yoy_q, 0.0, 0.0)),
    ]
    for metric_name, delta_txt, direction_txt in trend_rows:
        ws.cell(row=panel_row, column=panel_col, value=metric_name)
        ws.cell(row=panel_row, column=panel_col + 1, value=delta_txt)
        ws.cell(row=panel_row, column=panel_col + 2, value=direction_txt)
        ws.cell(row=panel_row, column=panel_col + 3, value=asof_txt)
        panel_row += 1
    row_trend_end = panel_row - 1

    panel_row += 1

    # Red/Green flags panel.
    row_flags_hdr = panel_row
    _clear_merges(row_flags_hdr, panel_col, panel_col + 8)
    ws.cell(row=panel_row, column=panel_col, value="Red/Green Flags").font = bold
    try:
        ws.merge_cells(start_row=panel_row, start_column=panel_col, end_row=panel_row, end_column=panel_col + 8)
    except Exception:
        pass
    for cc in range(panel_col, panel_col + 9):
        ws.cell(row=panel_row, column=cc).fill = section_fill
    panel_row += 1
    for cc in range(panel_col, panel_col + 9):
        ws.cell(row=panel_row, column=cc).fill = header_fill
    _clear_merges(panel_row, panel_col, panel_col + 8)
    hc = ws.cell(row=panel_row, column=panel_col, value="Flag")
    hc.font = bold
    hc.alignment = Alignment(horizontal="center", vertical="center")
    hc = ws.cell(row=panel_row, column=panel_col + 1, value="Status")
    hc.font = bold
    hc.alignment = Alignment(horizontal="center", vertical="center")
    hc = ws.cell(row=panel_row, column=panel_col + 2, value="Evidence")
    hc.font = bold
    hc.alignment = Alignment(horizontal="center", vertical="center")
    try:
        ws.merge_cells(start_row=panel_row, start_column=panel_col + 2, end_row=panel_row, end_column=panel_col + 7)
    except Exception:
        pass
    hc = ws.cell(row=panel_row, column=panel_col + 8, value="As-of")
    hc.font = bold
    hc.alignment = Alignment(horizontal="center", vertical="center")
    panel_row += 1

    def _flag_status(val: Optional[float], warn_thr: Optional[float], fail_thr: Optional[float], higher_worse: bool = True) -> str:
        if val is None or pd.isna(val):
            return "N/A"
        if warn_thr is None or fail_thr is None:
            return "PASS"
        if higher_worse:
            if val >= fail_thr:
                return "FAIL"
            if val >= warn_thr:
                return "WARN"
            return "PASS"
        if val <= fail_thr:
            return "FAIL"
        if val <= warn_thr:
            return "WARN"
        return "PASS"

    def _add_flag(flag_name: str, status: str, evidence: str) -> None:
        nonlocal panel_row
        _clear_merges(panel_row, panel_col, panel_col + 8)
        ws.cell(row=panel_row, column=panel_col, value=flag_name)
        st_cell = ws.cell(row=panel_row, column=panel_col + 1, value=status)
        st_cell.fill = _status_fill(status)
        ev_cell = ws.cell(row=panel_row, column=panel_col + 2, value=evidence)
        ev_cell.alignment = Alignment(wrap_text=True, vertical="top")
        try:
            ws.merge_cells(start_row=panel_row, start_column=panel_col + 2, end_row=panel_row, end_column=panel_col + 7)
        except Exception:
            pass
        ws.cell(row=panel_row, column=panel_col + 8, value=asof_txt)
        panel_row += 1

    cfo_ttm_q = _ttm_map(cfo_map).get(q0) if q0 is not None else None
    fcf_ttm_q = fcf_ttm_map.get(q0) if q0 is not None else None
    rev_yoy_q = _yoy_pct(rev_map, q0)
    cfo_yoy_q = _yoy_pct(cfo_map, q0)
    ar_yoy_q = _yoy_pct(ar_map, q0)
    inv_yoy_q = _yoy_pct(inventory_map, q0)
    debt_src_map = total_debt_map if any(v is not None for v in total_debt_map.values()) else debt_core_map
    debt_yoy_q = _yoy_pct(debt_src_map, q0)
    liq_yoy_q = _yoy_pct(liquidity_map, q0)
    cfo_net_income_ratio = (float(cfo_ttm_q) / float(net_income_ttm_map.get(q0))) if (cfo_ttm_q is not None and net_income_ttm_map.get(q0) not in (None, 0)) else None
    net_lev_yoy_delta = None
    if q0 is not None:
        prev_q0 = _anf_prior_year_quarter(q0, all_qs_ts) if is_anf_profile else pd.Timestamp(q0) - pd.DateOffset(years=1)
        if prev_q0 is not None and net_lev_map.get(q0) is not None and net_lev_map.get(prev_q0) is not None:
            net_lev_yoy_delta = float(net_lev_map.get(q0)) - float(net_lev_map.get(prev_q0))
    cov_cash_q = cov_cash_map.get(q0) if q0 is not None else None
    cov_cash_prev_q = _anf_prior_year_quarter(q0, all_qs_ts) if (q0 is not None and is_anf_profile) else (pd.Timestamp(q0) - pd.DateOffset(years=1) if q0 is not None else None)
    cov_cash_ly = cov_cash_map.get(cov_cash_prev_q) if cov_cash_prev_q is not None else None
    cov_cash_delta = (float(cov_cash_q) - float(cov_cash_ly)) if (cov_cash_q is not None and cov_cash_ly is not None) else None
    cap_returns_ttm = None
    if q0 is not None:
        buyback_ttm_q = buyback_ttm_map.get(q0)
        dividend_ttm_q = dividend_ttm_map.get(q0)
        if buyback_ttm_q is not None:
            cap_returns_ttm = float(buyback_ttm_q) + (float(dividend_ttm_q) if dividend_ttm_q is not None else 0.0)
        elif dividend_ttm_q is not None:
            cap_returns_ttm = float(dividend_ttm_q)
    equity_q = total_equity_map.get(q0) if q0 is not None else None
    pension_ratio = (float(pension_map.get(q0)) / abs(float(equity_q))) if (q0 is not None and pension_map.get(q0) is not None and equity_q not in (None, 0)) else None
    goodwill_q = goodwill_map.get(q0) if q0 is not None else None
    assets_q = assets_map.get(q0) if q0 is not None else None
    goodwill_pct_q = (goodwill_q / assets_q) if (goodwill_q is not None and assets_q not in (None, 0)) else None

    _add_flag("Red: Revenue up but CFO down (YoY)", "FAIL" if (rev_yoy_q is not None and cfo_yoy_q is not None and rev_yoy_q > 0.05 and cfo_yoy_q < -0.05) else ("PASS" if (rev_yoy_q is not None and cfo_yoy_q is not None) else "N/A"), f"Rev YoY {_pct(rev_yoy_q)} vs CFO YoY {_pct(cfo_yoy_q)}")
    _add_flag("Red: Earnings quality CFO/NI (TTM)", _flag_status(cfo_net_income_ratio, 1.0, 0.7, higher_worse=False), f"CFO/NI(TTM) {cfo_net_income_ratio:.2f}x" if cfo_net_income_ratio is not None else "N/A")
    ar_gap = (ar_yoy_q - rev_yoy_q) if (ar_yoy_q is not None and rev_yoy_q is not None) else None
    _add_flag("Red: AR growing faster than revenue (YoY)", _flag_status(ar_gap, 0.10, 0.20, higher_worse=True), f"AR YoY {_pct(ar_yoy_q)} vs Rev YoY {_pct(rev_yoy_q)} (gap {_pct(ar_gap)})")
    _add_flag("Red: Inventory build without revenue growth", "WARN" if (inv_yoy_q is not None and rev_yoy_q is not None and inv_yoy_q > 0.10 and rev_yoy_q < 0.03) else ("PASS" if (inv_yoy_q is not None and rev_yoy_q is not None) else "N/A"), f"Inv YoY {_pct(inv_yoy_q)} vs Rev YoY {_pct(rev_yoy_q)}")
    debt_gap = (debt_yoy_q - rev_yoy_q) if (debt_yoy_q is not None and rev_yoy_q is not None) else None
    _add_flag("Red: Debt growing faster than revenue (YoY)", _flag_status(debt_gap, 0.10, 0.20, higher_worse=True), f"Debt YoY {_pct(debt_yoy_q)} vs Rev YoY {_pct(rev_yoy_q)} (gap {_pct(debt_gap)})")
    _add_flag("Red: Leverage rising (YoY Δ)", _flag_status(net_lev_yoy_delta, 0.5, 1.0, higher_worse=True), f"Net leverage YoY Δ {net_lev_yoy_delta:.2f}x" if net_lev_yoy_delta is not None else "N/A")
    _add_flag("Red: Interest coverage low (cash)", _flag_status(cov_cash_q, 2.0, 1.0, higher_worse=False), f"Coverage cash {cov_cash_q:.2f}x" if cov_cash_q is not None else "N/A")
    _add_flag("Red: FCF negative while EBITDA positive (TTM)", "WARN" if (fcf_ttm_q is not None and ebitda_ttm_map.get(q0) is not None and float(fcf_ttm_q) < 0 and float(ebitda_ttm_map.get(q0)) > 0) else ("PASS" if (fcf_ttm_q is not None and ebitda_ttm_map.get(q0) is not None) else "N/A"), f"FCF TTM {_money_m(fcf_ttm_q)} vs EBITDA TTM {_money_m(ebitda_ttm_map.get(q0) if q0 is not None else None)}")
    cap_ret_status = "N/A"
    if cap_returns_ttm is not None and fcf_ttm_q is not None and net_debt_yoy_delta_q is not None:
        if cap_returns_ttm > fcf_ttm_q and net_debt_yoy_delta_q > 0:
            if q0 is not None and net_debt_map.get(q0) is not None and float(net_debt_map.get(q0)) < 0:
                cap_ret_status = "WARN"
            else:
                cap_ret_status = "FAIL" if cap_returns_ttm > fcf_ttm_q * 1.2 else "WARN"
        else:
            cap_ret_status = "PASS"
    if (
        q0 is not None
        and cap_returns_ttm is not None
        and fcf_ttm_q is not None
        and cap_returns_ttm > fcf_ttm_q
        and net_debt_map.get(q0) is not None
        and float(net_debt_map.get(q0)) < 0
    ):
        _add_flag(
            "Watch: Buybacks exceeded FCF",
            "WARN",
            f"Buybacks/returns TTM {_money_m(cap_returns_ttm)} vs FCF TTM {_money_m(fcf_ttm_q)}; funded partly from net cash.",
        )
    else:
        _add_flag(
            "Red: Capital returns exceed FCF while net debt rising",
            cap_ret_status,
            f"Returns TTM {_money_m(cap_returns_ttm)} vs FCF TTM {_money_m(fcf_ttm_q)}; Net debt YoY Δ {_delta_m(net_debt_yoy_delta_q)}",
        )
    _add_flag("Red: Goodwill heavy", _flag_status(goodwill_pct_q, 0.30, 0.50, higher_worse=True), f"Goodwill/assets {_pct(goodwill_pct_q)}")
    _add_flag("Red: Share dilution (YoY)", _flag_status(shares_yoy_q, 0.03, 0.07, higher_worse=True), f"Shares YoY {_pct(shares_yoy_q)}")
    _add_flag("Red: Pension obligations pressure", _flag_status(pension_ratio, 0.5, 1.0, higher_worse=True), f"Pension/|equity| {pension_ratio:.2f}x" if pension_ratio is not None else "N/A")
    panel_row += 1  # blank spacer between red and green

    op_qoq_pp = None
    op_margin_map = _margin(ebit_map, rev_map)
    if len(last4_qs) >= 2 and op_margin_map.get(last4_qs[-1]) is not None and op_margin_map.get(last4_qs[-2]) is not None:
        op_qoq_pp = (float(op_margin_map.get(last4_qs[-1])) - float(op_margin_map.get(last4_qs[-2]))) * 100.0
    if op_qoq_pp is None:
        op_qoq_status = "N/A"
    elif op_qoq_pp > 0.5:
        op_qoq_status = "PASS"
    elif op_qoq_pp < -0.5:
        op_qoq_status = "FAIL"
    else:
        op_qoq_status = "WARN"
    _add_flag("Green: Operating margin trend QoQ", op_qoq_status, f"Operating margin QoQ Δ {op_qoq_pp:.2f}pp" if op_qoq_pp is not None else "N/A")
    _add_flag("Green: FCF TTM growth (YoY)", "PASS" if (fcf_ttm_y is not None and fcf_ttm_y > 0.10) else ("WARN" if (fcf_ttm_y is not None and fcf_ttm_y >= 0) else ("FAIL" if fcf_ttm_y is not None else "N/A")), f"FCF TTM YoY {_pct(fcf_ttm_y)}")
    net_debt_current_q = net_debt_map.get(q0) if q0 is not None else None
    net_debt_flag_name, net_debt_flag_status = _net_debt_yoy_flag_label_and_status_for_position(
        net_debt_yoy_delta_q,
        net_debt_current_q,
    )
    if net_debt_current_q is not None and net_debt_yoy_delta_q is not None and float(net_debt_current_q) < 0 and float(net_debt_yoy_delta_q) > 0:
        net_debt_evidence = f"Net cash decreased by {_money_m(abs(net_debt_yoy_delta_q))} YoY; still net cash {_money_m(abs(net_debt_current_q))}"
    else:
        net_debt_evidence = f"Net debt YoY Δ {_delta_m(net_debt_yoy_delta_q)}"
    _add_flag(net_debt_flag_name, net_debt_flag_status, net_debt_evidence)
    _add_flag("Green: Interest coverage improving (YoY)", "PASS" if (cov_cash_delta is not None and cov_cash_delta > 0.5) else ("WARN" if (cov_cash_delta is not None and cov_cash_delta >= 0) else ("FAIL" if cov_cash_delta is not None else "N/A")), f"Coverage cash YoY Δ {cov_cash_delta:.2f}x" if cov_cash_delta is not None else "N/A")
    _add_flag("Green: Shares outstanding decreasing (YoY)", "PASS" if (shares_yoy_q is not None and shares_yoy_q < 0) else ("FAIL" if shares_yoy_q is not None else "N/A"), f"Shares YoY {_pct(shares_yoy_q)}")
    _add_flag("Green: Liquidity improving (YoY)", "PASS" if (liq_yoy_q is not None and liq_yoy_q > 0) else ("WARN" if (liq_yoy_q is not None and abs(float(liq_yoy_q)) <= 0.01) else ("FAIL" if liq_yoy_q is not None else "N/A")), f"Liquidity YoY {_pct(liq_yoy_q)}")
    row_flags_end = panel_row - 1

    return ValuationTrendFlagsRenderResult(
        row_trend_hdr=row_trend_hdr,
        row_trend_end=row_trend_end,
        row_flags_hdr=row_flags_hdr,
        row_flags_end=row_flags_end,
        panel_col=panel_col,
        panel_row_start=panel_row_start,
        next_panel_row=panel_row,
    )
