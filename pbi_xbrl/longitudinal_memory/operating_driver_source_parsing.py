"""Shared, fail-closed parsers for Operating Drivers primary-source evidence.

The helpers in this module know nothing about tickers or workbook coordinates.
They expose typed source observations from official spreadsheet, inline-XBRL,
HTML-table, and narrative evidence.  Economic mapping remains declarative in a
ticker or sector profile, and every parser preserves missing source cells.
"""
from __future__ import annotations

from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
from typing import Any, Mapping, Sequence

from lxml import etree, html
from openpyxl import load_workbook


class OperatingDriverSourceParsingError(ValueError):
    """Raised when source evidence cannot be parsed without ambiguity."""


@dataclass(frozen=True)
class QuarterlyTableObservation:
    metric_key: str
    source_label: str
    fiscal_year: int
    fiscal_quarter: int
    value: Decimal | None
    source_cell: str
    source_state: str

    def to_dict(self) -> dict[str, Any]:
        result = asdict(self)
        result["value"] = None if self.value is None else _decimal_text(self.value)
        return result


@dataclass(frozen=True)
class InlineXbrlInstantFact:
    concept_name: str
    context_ref: str
    instant_date: str
    value: Decimal
    unit_ref: str | None
    source_text: str

    def to_dict(self) -> dict[str, Any]:
        result = asdict(self)
        result["value"] = _decimal_text(self.value)
        return result


@dataclass(frozen=True)
class RetailActivitySnapshot:
    new_stores: int | None
    remodeled_stores: int | None
    right_sized_stores: int | None
    closed_stores: int | None
    matched_text: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class AdditiveQuarterResult:
    fiscal_year: int
    fiscal_quarter: int
    value: Decimal
    current_cumulative_value: Decimal
    prior_cumulative_value: Decimal | None
    rule: str = "current-ytd-actual-minus-prior-ytd-actual"

    def to_dict(self) -> dict[str, Any]:
        result = asdict(self)
        for key in ("value", "current_cumulative_value", "prior_cumulative_value"):
            value = result[key]
            result[key] = None if value is None else _decimal_text(value)
        return result


_WORD_NUMBERS = {
    "zero": 0,
    "one": 1,
    "two": 2,
    "three": 3,
    "four": 4,
    "five": 5,
    "six": 6,
    "seven": 7,
    "eight": 8,
    "nine": 9,
    "ten": 10,
    "eleven": 11,
    "twelve": 12,
    "thirteen": 13,
    "fourteen": 14,
    "fifteen": 15,
    "sixteen": 16,
    "seventeen": 17,
    "eighteen": 18,
    "nineteen": 19,
    "twenty": 20,
    "twenty-one": 21,
    "twenty-two": 22,
    "twenty-three": 23,
    "twenty-four": 24,
    "twenty-five": 25,
    "twenty-six": 26,
    "twenty-seven": 27,
    "twenty-eight": 28,
    "twenty-nine": 29,
    "thirty": 30,
    "thirty-five": 35,
    "forty": 40,
    "forty-one": 41,
    "forty-seven": 47,
    "forty-eight": 48,
    "fifty": 50,
    "fifty-nine": 59,
    "sixty": 60,
    "sixty-two": 62,
    "sixty-five": 65,
    "seventy": 70,
    "eighty": 80,
}
_COUNT_TOKEN = r"(?:\d{1,4}|[a-z]+(?:-[a-z]+)?)"


def _decimal_text(value: Decimal) -> str:
    result = format(value, "f")
    if "." in result:
        result = result.rstrip("0").rstrip(".")
    return result or "0"


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", "" if value is None else str(value)).strip()


def _normal_label(value: Any) -> str:
    text = _clean(value).casefold()
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _source_decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = _clean(value)
    if not text or text.casefold() in {"not provided", "n/a", "na", "—", "-"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()$,% ").replace(",", "")
    try:
        result = Decimal(text)
    except InvalidOperation as exc:
        raise OperatingDriverSourceParsingError(f"Unparseable numeric source cell: {value!r}") from exc
    return -result if negative else result


def parse_quarterly_history_table(
    path: Path | str,
    *,
    sheet_name: str,
    metric_aliases: Mapping[str, Sequence[str]],
    fiscal_group_token: str = "Fiscal",
) -> tuple[QuarterlyTableObservation, ...]:
    """Read fiscal-quarter observations from an issuer history table.

    The parser discovers fiscal-year group headers and Q1-Q4 columns rather
    than owning specific cells.  Row identity is supplied by a declarative
    alias map.  Explicit "Not provided" cells are returned, never discarded.
    """

    source = Path(path)
    if not source.is_file():
        raise OperatingDriverSourceParsingError(f"Source workbook is absent: {source}")
    workbook = load_workbook(source, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise OperatingDriverSourceParsingError(
                f"Source sheet {sheet_name!r} is absent from {source.name}."
            )
        sheet = workbook[sheet_name]
        fiscal_row: int | None = None
        quarter_row: int | None = None
        for row in range(1, min(sheet.max_row, 20) + 1):
            values = [_clean(sheet.cell(row, column).value) for column in range(1, sheet.max_column + 1)]
            if any(value.startswith(f"{fiscal_group_token} ") for value in values):
                fiscal_row = row
                for candidate in range(row + 1, min(row + 4, sheet.max_row) + 1):
                    candidate_values = {
                        _clean(sheet.cell(candidate, column).value).upper()
                        for column in range(1, sheet.max_column + 1)
                    }
                    if candidate_values & {"Q1", "Q2", "Q3", "Q4"}:
                        quarter_row = candidate
                        break
                if quarter_row is not None:
                    break
        if fiscal_row is None or quarter_row is None:
            raise OperatingDriverSourceParsingError(
                f"Could not discover fiscal/quarter headers in {source.name}:{sheet_name}."
            )

        periods: dict[int, tuple[int, int]] = {}
        active_year: int | None = None
        for column in range(1, sheet.max_column + 1):
            group = _clean(sheet.cell(fiscal_row, column).value)
            match = re.search(r"\b(?:Fiscal\s+)?(20\d{2})\b", group, re.I)
            if match:
                active_year = int(match.group(1))
            quarter = _clean(sheet.cell(quarter_row, column).value).upper()
            if active_year is not None and quarter in {"Q1", "Q2", "Q3", "Q4"}:
                periods[column] = (active_year, int(quarter[1]))

        alias_lookup: dict[str, str] = {}
        for metric_key, aliases in metric_aliases.items():
            for alias in aliases:
                normalized = _normal_label(alias)
                if normalized in alias_lookup and alias_lookup[normalized] != metric_key:
                    raise OperatingDriverSourceParsingError(
                        f"Ambiguous declarative alias {alias!r}."
                    )
                alias_lookup[normalized] = metric_key

        found_rows: dict[str, int] = {}
        equivalent_duplicate_rows: dict[str, list[int]] = {}
        labels: dict[str, str] = {}
        for row in range(quarter_row + 1, sheet.max_row + 1):
            label = _clean(sheet.cell(row, 1).value)
            normalized = _normal_label(label)
            metric_key = alias_lookup.get(normalized)
            if metric_key is None:
                continue
            if metric_key in found_rows:
                prior_row = found_rows[metric_key]
                prior_values = tuple(
                    _clean(sheet.cell(prior_row, column).value)
                    for column in sorted(periods)
                )
                candidate_values = tuple(
                    _clean(sheet.cell(row, column).value)
                    for column in sorted(periods)
                )
                if candidate_values != prior_values:
                    raise OperatingDriverSourceParsingError(
                        f"Conflicting duplicate rows matched {metric_key!r} in "
                        f"{source.name}:{sheet_name}."
                    )
                equivalent_duplicate_rows.setdefault(metric_key, []).append(row)
                continue
            found_rows[metric_key] = row
            labels[metric_key] = label

        missing_metrics = sorted(set(metric_aliases) - set(found_rows))
        if missing_metrics:
            raise OperatingDriverSourceParsingError(
                f"Declarative metrics absent from {source.name}:{sheet_name}: {missing_metrics}"
            )

        result: list[QuarterlyTableObservation] = []
        for metric_key in sorted(found_rows):
            row = found_rows[metric_key]
            for column, (year, quarter) in sorted(periods.items()):
                raw = sheet.cell(row, column).value
                value = _source_decimal(raw)
                raw_text = _clean(raw)
                state = (
                    "DIRECT_NUMERIC"
                    if value is not None
                    else "NOT_DISCLOSED"
                    if raw_text.casefold() == "not provided"
                    else "EMPTY_SOURCE_CELL"
                )
                result.append(
                    QuarterlyTableObservation(
                        metric_key=metric_key,
                        source_label=labels[metric_key],
                        fiscal_year=year,
                        fiscal_quarter=quarter,
                        value=value,
                        source_cell=sheet.cell(row, column).coordinate,
                        source_state=state,
                    )
                )
        return tuple(result)
    finally:
        workbook.close()


def parse_inline_xbrl_instant_facts(
    path: Path | str,
    *,
    concept_names: Sequence[str],
) -> tuple[InlineXbrlInstantFact, ...]:
    """Extract dimensionless inline-XBRL instant facts from an SEC filing."""

    source = Path(path)
    if not source.is_file():
        raise OperatingDriverSourceParsingError(f"SEC filing is absent: {source}")
    document = etree.parse(
        str(source),
        etree.XMLParser(huge_tree=True, recover=False, resolve_entities=False),
    )
    contexts: dict[str, tuple[str, bool]] = {}
    for context in document.xpath("//*[local-name()='context']"):
        context_id = context.attrib.get("id")
        instant = context.xpath(".//*[local-name()='period']/*[local-name()='instant']/text()")
        if context_id and instant:
            has_dimensions = bool(context.xpath(".//*[local-name()='segment']/*"))
            contexts[context_id] = (_clean(instant[0]), has_dimensions)

    allowed = set(concept_names)
    result: list[InlineXbrlInstantFact] = []
    seen: set[tuple[str, str, Decimal, str | None]] = set()
    for element in document.xpath("//*[@name]"):
        name = element.attrib.get("name")
        context_ref = element.attrib.get("contextRef") or element.attrib.get("contextref")
        if name not in allowed or context_ref not in contexts:
            continue
        instant_date, has_dimensions = contexts[context_ref]
        if has_dimensions:
            continue
        raw = _clean(" ".join(element.itertext()))
        numeric = _source_decimal(raw)
        if numeric is None:
            continue
        scale_text = element.attrib.get("scale")
        scale = int(scale_text) if scale_text else 0
        value = numeric * (Decimal(10) ** scale)
        if element.attrib.get("sign") == "-":
            value = -value
        unit_ref = element.attrib.get("unitRef") or element.attrib.get("unitref")
        key = (name, instant_date, value, unit_ref)
        if key in seen:
            continue
        seen.add(key)
        result.append(
            InlineXbrlInstantFact(
                concept_name=name,
                context_ref=context_ref,
                instant_date=instant_date,
                value=value,
                unit_ref=unit_ref,
                source_text=raw,
            )
        )
    return tuple(sorted(result, key=lambda item: (item.concept_name, item.instant_date, item.value)))


def parse_html_table_terminal_number(
    path: Path | str,
    *,
    required_table_text: str,
    row_label: str,
    section_label: str | None = None,
) -> Decimal:
    """Read the final numeric cell from one unambiguous HTML table row.

    ``section_label`` is a declarative disambiguator for filings that repeat a
    date row in adjacent store-count and gross-square-footage sections.
    """

    source = Path(path)
    document = html.fromstring(source.read_bytes())
    candidates = []
    for table in document.xpath("//table"):
        table_text = _clean(" ".join(table.itertext()))
        if required_table_text.casefold() not in table_text.casefold():
            continue
        in_section = section_label is None
        for row in table.xpath(".//tr"):
            cells = [_clean(" ".join(cell.itertext())) for cell in row.xpath("./th|./td")]
            first = cells[0] if cells else ""
            if section_label is not None and first.casefold() == section_label.casefold():
                in_section = True
                continue
            if (
                section_label is not None
                and in_section
                and first.endswith(":")
                and first.casefold() != section_label.casefold()
            ):
                in_section = False
            if in_section and first.casefold() == row_label.casefold():
                numeric = [_source_decimal(cell) for cell in cells[1:]]
                numeric = [value for value in numeric if value is not None]
                if numeric:
                    candidates.append(numeric[-1])
    unique_candidates = tuple(dict.fromkeys(candidates))
    if len(unique_candidates) != 1:
        raise OperatingDriverSourceParsingError(
            f"Expected one semantic value for {row_label!r} in a "
            f"{required_table_text!r} table; found {len(unique_candidates)}."
        )
    return unique_candidates[0]


def _count(token: str | None) -> int | None:
    if token is None:
        return None
    normalized = token.casefold().strip().replace(" ", "-")
    if normalized.isdigit():
        return int(normalized)
    if normalized not in _WORD_NUMBERS:
        raise OperatingDriverSourceParsingError(f"Unsupported count token: {token!r}")
    return _WORD_NUMBERS[normalized]


def parse_retail_activity_snapshot(text: str) -> RetailActivitySnapshot:
    """Parse one issuer sentence containing cumulative or annual store activity."""

    cleaned = _clean(text).replace("right sized", "right-sized")
    patterns = (
        re.compile(
            rf"opened\s+(?P<new>{_COUNT_TOKEN})\s+new store(?:s| locations)?\s*,?\s*"
            rf"remodeled\s+(?P<remodel>{_COUNT_TOKEN})\s+store(?:s| locations)?\s*,?\s*(?:and\s*)?"
            rf"right-sized\s+(?:an additional\s+)?(?P<right>{_COUNT_TOKEN})\s+store(?:s| locations)?\s*,?\s*"
            rf"(?:(?:and|while)\s+)?clos(?:ed|ing)\s+(?P<closed>{_COUNT_TOKEN})\s+stores?",
            re.I,
        ),
        re.compile(
            rf"opened\s+(?P<new>{_COUNT_TOKEN})\s+new stores?\s*,?\s*while\s+closing\s+"
            rf"(?P<closed>{_COUNT_TOKEN})\s+stores?",
            re.I,
        ),
    )
    matches = [pattern.search(cleaned) for pattern in patterns]
    matches = [match for match in matches if match is not None]
    if not matches:
        raise OperatingDriverSourceParsingError("No supported store-activity statement was found.")
    match = matches[0]
    return RetailActivitySnapshot(
        new_stores=_count(match.groupdict().get("new")),
        remodeled_stores=_count(match.groupdict().get("remodel")),
        right_sized_stores=_count(match.groupdict().get("right")),
        closed_stores=_count(match.groupdict().get("closed")),
        matched_text=match.group(0),
    )


def derive_additive_quarter_actuals(
    *,
    fiscal_year: int,
    cumulative_actuals: Mapping[int, Decimal | int | str],
) -> tuple[AdditiveQuarterResult, ...]:
    """Difference adjacent compatible cumulative actuals without gap bridging."""

    normalized = {quarter: Decimal(str(value)) for quarter, value in cumulative_actuals.items()}
    if any(quarter not in {1, 2, 3, 4} for quarter in normalized):
        raise OperatingDriverSourceParsingError("Fiscal-quarter keys must be between 1 and 4.")
    result: list[AdditiveQuarterResult] = []
    for quarter in range(1, 5):
        if quarter not in normalized:
            continue
        if quarter > 1 and quarter - 1 not in normalized:
            continue
        prior = None if quarter == 1 else normalized[quarter - 1]
        value = normalized[quarter] if prior is None else normalized[quarter] - prior
        result.append(
            AdditiveQuarterResult(
                fiscal_year=fiscal_year,
                fiscal_quarter=quarter,
                value=value,
                current_cumulative_value=normalized[quarter],
                prior_cumulative_value=prior,
            )
        )
    return tuple(result)


__all__ = [
    "AdditiveQuarterResult",
    "InlineXbrlInstantFact",
    "OperatingDriverSourceParsingError",
    "QuarterlyTableObservation",
    "RetailActivitySnapshot",
    "derive_additive_quarter_actuals",
    "parse_html_table_terminal_number",
    "parse_inline_xbrl_instant_facts",
    "parse_quarterly_history_table",
    "parse_retail_activity_snapshot",
]
