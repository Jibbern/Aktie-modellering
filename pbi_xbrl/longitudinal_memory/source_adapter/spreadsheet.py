"""Read-only XLSX cell extraction with formula and format replay."""
from __future__ import annotations

import hashlib
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Mapping

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from pbi_xbrl.longitudinal_memory.types import canonical_decimal

from .types import DiscoveredDocument, ExtractedEvidence, LocatorError, text_sha256


METHOD_ID = "extractor:source:xlsx-read-only-cell@1"


def _text(value: Any) -> str:
    return " ".join(str(value or "").split())


def _range_values(sheet: Any, a1_range: str) -> list[Any]:
    min_col, min_row, max_col, max_row = range_boundaries(a1_range)
    return [
        _display_header_value(sheet.cell(row=row, column=column).value)
        for row in range(min_row, max_row + 1)
        for column in range(min_col, max_col + 1)
        if sheet.cell(row=row, column=column).value is not None
    ]


def _display_header_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return f"{value.strftime('%B')} {value.day}, {value.year}"
    return value


def _canonical_cell_value(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, Decimal)):
        return canonical_decimal(value)
    if isinstance(value, float):
        return canonical_decimal(Decimal(str(value)))
    return _text(value)


def extract_spreadsheet_evidence(
    document: DiscoveredDocument,
    assertions: list[Mapping[str, Any]],
) -> tuple[ExtractedEvidence, ...]:
    formula_stream = BytesIO(document.verified_bytes)
    cached_stream = BytesIO(document.verified_bytes)
    if (
        hashlib.sha256(formula_stream.getbuffer()).hexdigest() != document.content_sha256
        or hashlib.sha256(cached_stream.getbuffer()).hexdigest() != document.content_sha256
    ):
        raise LocatorError(f"XLSX byte snapshot changed for {document.spec.document_key!r}.")
    formulas = load_workbook(formula_stream, read_only=True, data_only=False)
    cached = load_workbook(cached_stream, read_only=True, data_only=True)
    try:
        result: list[ExtractedEvidence] = []
        for assertion in sorted(assertions, key=lambda row: str(row["assertion_key"])):
            locator = assertion["locator"]
            if locator["locator_kind"] not in {"xlsx-cell", "xlsx-range"}:
                raise LocatorError(f"Unsupported XLSX locator kind {locator['locator_kind']!r}.")
            if locator["extraction_method_id"] != METHOD_ID:
                raise LocatorError(
                    f"XLSX extraction method changed for {assertion['assertion_key']!r}."
                )
            sheet_name = str(locator["sheet_name"])
            if sheet_name not in formulas.sheetnames:
                raise LocatorError(f"XLSX sheet {sheet_name!r} does not exist.")
            formula_sheet, cached_sheet = formulas[sheet_name], cached[sheet_name]
            header_values = [
                value
                for header_range in locator["header_ranges"]
                for value in _range_values(formula_sheet, str(header_range))
            ]
            header_text = " | ".join(_text(value) for value in header_values if _text(value))
            if _text(locator["row_header_fingerprint"]).casefold() not in header_text.casefold():
                raise LocatorError(f"XLSX row fingerprint changed for {assertion['assertion_key']!r}.")
            column_fingerprints = [
                _text(value).casefold()
                for value in str(locator["column_header_fingerprint"]).split("|")
                if _text(value)
            ]
            if not column_fingerprints or not all(
                value in header_text.casefold() for value in column_fingerprints
            ):
                raise LocatorError(f"XLSX column fingerprint changed for {assertion['assertion_key']!r}.")

            min_col, min_row, max_col, max_row = range_boundaries(str(locator["a1_range"]))
            if locator["locator_kind"] == "xlsx-cell" and (min_col, min_row) != (max_col, max_row):
                raise LocatorError("An xlsx-cell locator must name exactly one cell.")
            cells = [
                formula_sheet.cell(row=row, column=column)
                for row in range(min_row, max_row + 1)
                for column in range(min_col, max_col + 1)
            ]
            if len(cells) != 1:
                raise LocatorError("The bounded first pass accepts one-cell XLSX evidence only.")
            cell = cells[0]
            cached_cell = cached_sheet.cell(row=min_row, column=min_col)
            actual_type = "formula" if cell.data_type == "f" else "numeric" if isinstance(cell.value, (int, float, Decimal)) and not isinstance(cell.value, bool) else "text"
            if actual_type != locator["cell_type"]:
                raise LocatorError(
                    f"XLSX cell type changed for {assertion['assertion_key']!r}: "
                    f"expected {locator['cell_type']!r}, received {actual_type!r}."
                )
            if cell.number_format != locator["number_format"]:
                raise LocatorError(f"XLSX number format changed for {assertion['assertion_key']!r}.")
            actual_formula = str(cell.value) if actual_type == "formula" else None
            if actual_formula != locator["formula"]:
                raise LocatorError(f"XLSX formula state changed for {assertion['assertion_key']!r}.")
            cached_state = "not-applicable" if actual_type != "formula" else "present" if cached_cell.value is not None else "missing"
            if cached_state != locator["cached_value_state"]:
                raise LocatorError(f"XLSX cached-value state changed for {assertion['assertion_key']!r}.")
            raw_value = cached_cell.value if actual_type == "formula" else cell.value
            value_text = _canonical_cell_value(raw_value)
            excerpt = (
                f"{sheet_name}!{locator['a1_range']} | {locator['row_header_fingerprint']} | "
                f"{locator['column_header_fingerprint']} | {value_text}"
            )
            if excerpt != locator["excerpt"]:
                raise LocatorError(
                    f"XLSX excerpt mismatch for {assertion['assertion_key']!r}: "
                    f"expected {locator['excerpt']!r}, received {excerpt!r}."
                )
            if text_sha256(excerpt) != locator["excerpt_sha256"]:
                raise LocatorError(f"XLSX excerpt digest mismatch for {assertion['assertion_key']!r}.")
            result.append(
                ExtractedEvidence(
                    assertion_key=str(assertion["assertion_key"]),
                    document_key=document.spec.document_key,
                    locator_kind="cell" if locator["locator_kind"] == "xlsx-cell" else "range",
                    locator_key=str(locator["locator_key"]),
                    ordinal=int(locator["ordinal"]),
                    extraction_method_id=str(locator["extraction_method_id"]),
                    excerpt=excerpt,
                    excerpt_sha256=str(locator["excerpt_sha256"]),
                    value_text=value_text,
                    comparison_text=None,
                    review_state=str(locator["review_state"]),
                    diagnostics={
                        "sheet_name": sheet_name,
                        "a1_range": locator["a1_range"],
                        "cell_type": actual_type,
                        "number_format": cell.number_format,
                        "formula": actual_formula,
                        "cached_value_state": cached_state,
                    },
                )
            )
        return tuple(result)
    finally:
        formulas.close()
        cached.close()
