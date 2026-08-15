"""Bounded Valuation correction and canonical Investment Case consumer plan.

The exhaustive audit remains the authority for affected historical cells and
legacy formula dispositions.  This module translates those accepted findings
into an immutable workbook mutation plan; it performs no source research.
"""
from __future__ import annotations

from dataclasses import asdict, dataclass
from decimal import Decimal
import hashlib
import json
from pathlib import Path
import re
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.utils import get_column_letter, range_boundaries

from pbi_xbrl.longitudinal_memory.formula_aware_workbook_materialization import (
    DefinedNameMutation,
    FormulaAwareCellMutation,
    WorkbookCalculationMetadataPolicy,
    WorksheetMergeMutation,
    WorksheetRowMutation,
)
from pbi_xbrl.new_ticker_investment_case_formula_surface import (
    CANONICAL_VALUATION_MATRIX_RANGE,
    canonical_investment_case_defined_names,
)


PROJECTION_CONTRACT = "anf-valuation-bounded-source-native-correction@1"
CALCULATION_METADATA_POLICY_ID = "valuation-native-safe-calculation-metadata@1"
VALUATION_CALCULATION_METADATA_POLICY = WorkbookCalculationMetadataPolicy(
    policy_id=CALCULATION_METADATA_POLICY_ID,
    expected_calc_mode="auto",
    expected_full_calc_on_load=True,
    expected_force_full_calc=True,
    force_full_calc=False,
)
AUDIT_SCHEMA = "valuation-exhaustive-reconciliation@1"
BASE_SUMMARY_BS_GOLDEN_SHA256 = "f57854d278b27bf206222d1979cba218d79aa355b5a36239f84af4950d6cbda2"
PROTECTED_ANF_SHA256 = "ef73bdef6b6efa1bc358622b58bfc320b609c128e76c602de9d4e5f726ab98cd"

IC_VISIBLE_SHEET = "ANF_Investment_Case"
IC_DATA_SHEET = "ANF_Investment_Case_Data"
VALUATION_SHEET = "Valuation"
ALLOWED_IC_DEPENDENCY_SHEETS = frozenset({IC_VISIBLE_SHEET, IC_DATA_SHEET})

_CELL_RE = re.compile(r"\$?([A-Z]{1,3})\$?([1-9][0-9]*)\Z")
_RANGE_RE = re.compile(
    r"\$?([A-Z]{1,3})\$?([1-9][0-9]*):\$?([A-Z]{1,3})\$?([1-9][0-9]*)\Z"
)


class ValuationProjectionError(ValueError):
    """Fail-closed projection-plan validation error."""


@dataclass(frozen=True)
class InvestmentCaseDependencyClosure:
    seed_count: int
    cell_count: int
    formula_count: int
    value_count: int
    blank_count: int
    cells_by_sheet: Mapping[str, int]
    cells: tuple[tuple[str, str], ...]


@dataclass(frozen=True)
class ValuationProjectionPlan:
    contract: str
    audit_manifest_sha256: str
    base_workbook_sha256: str
    donor_workbook_sha256: str
    cell_mutations: tuple[FormulaAwareCellMutation, ...]
    defined_name_mutations: tuple[DefinedNameMutation, ...]
    merge_mutations: tuple[WorksheetMergeMutation, ...]
    row_mutations: tuple[WorksheetRowMutation, ...]
    ic_dependency_closure: InvestmentCaseDependencyClosure
    liquidity_issue_cells: tuple[str, ...]
    securities_net_cash_issue_cells: tuple[str, ...]
    interest_coverage_cells: tuple[str, ...]
    old_formula_retirement_cells: tuple[str, ...]
    compact_link_cells: tuple[str, ...]
    legacy_name_deletions: tuple[str, ...]
    legacy_name_rebindings: Mapping[str, str]
    projection_digest: str
    formula_plan_digest: str
    defined_name_plan_digest: str

    def as_dict(self, *, include_mutations: bool = False) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "audit_manifest_sha256": self.audit_manifest_sha256,
            "base_workbook_sha256": self.base_workbook_sha256,
            "calculation_metadata_policy_id": CALCULATION_METADATA_POLICY_ID,
            "cell_mutation_count": len(self.cell_mutations),
            "compact_link_cells": list(self.compact_link_cells),
            "contract": self.contract,
            "defined_name_mutation_count": len(self.defined_name_mutations),
            "defined_name_plan_digest": self.defined_name_plan_digest,
            "donor_workbook_sha256": self.donor_workbook_sha256,
            "formula_plan_digest": self.formula_plan_digest,
            "ic_dependency_closure": asdict(self.ic_dependency_closure),
            "interest_coverage_cells": list(self.interest_coverage_cells),
            "legacy_name_deletions": list(self.legacy_name_deletions),
            "legacy_name_rebindings": dict(self.legacy_name_rebindings),
            "merge_mutations": [asdict(item) for item in self.merge_mutations],
            "row_mutations": [asdict(item) for item in self.row_mutations],
            "liquidity_issue_cells": list(self.liquidity_issue_cells),
            "old_formula_retirement_cells": list(self.old_formula_retirement_cells),
            "projection_digest": self.projection_digest,
            "securities_net_cash_issue_cells": list(self.securities_net_cash_issue_cells),
        }
        if include_mutations:
            payload["cell_mutations"] = [asdict(item) for item in self.cell_mutations]
            payload["defined_name_mutations"] = [asdict(item) for item in self.defined_name_mutations]
        return payload


def _canonical_bytes(value: Any) -> bytes:
    return (json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")) + "\n").encode(
        "utf-8"
    )


def _digest(value: Any) -> str:
    return hashlib.sha256(_canonical_bytes(value)).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _load_json(path: Path) -> dict[str, Any]:
    def reject_duplicates(pairs: Sequence[tuple[str, Any]]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for key, value in pairs:
            if key in result:
                raise ValuationProjectionError(f"Duplicate JSON key {key!r} in {path}.")
            result[key] = value
        return result

    with path.open("r", encoding="utf-8") as handle:
        value = json.load(handle, object_pairs_hook=reject_duplicates)
    if not isinstance(value, dict):
        raise ValuationProjectionError(f"Expected an object in {path}.")
    return value


def _cell_key(sheet: str, coordinate: str) -> tuple[str, str]:
    normalized = coordinate.replace("$", "").upper()
    if _CELL_RE.fullmatch(normalized) is None:
        raise ValuationProjectionError(f"Invalid cell reference {sheet}!{coordinate}.")
    return sheet, normalized


def _expand_reference(current_sheet: str, reference: str) -> tuple[tuple[str, str], ...] | None:
    value = reference.strip()
    sheet = current_sheet
    if "!" in value:
        sheet_text, value = value.rsplit("!", 1)
        sheet = sheet_text.strip("'").replace("''", "'")
    value = value.replace("$", "")
    if _CELL_RE.fullmatch(value):
        return (_cell_key(sheet, value),)
    range_match = _RANGE_RE.fullmatch(value)
    if range_match is None:
        return None
    minimum_column, minimum_row, maximum_column, maximum_row = range_boundaries(value)
    size = (maximum_column - minimum_column + 1) * (maximum_row - minimum_row + 1)
    if size > 5000:
        raise ValuationProjectionError(f"Dependency range is unexpectedly broad: {reference!r}.")
    return tuple(
        (sheet, f"{get_column_letter(column)}{row}")
        for row in range(minimum_row, maximum_row + 1)
        for column in range(minimum_column, maximum_column + 1)
    )


def _formula_dependencies(sheet: str, coordinate: str, formula: str) -> tuple[tuple[str, str], ...]:
    result: set[tuple[str, str]] = set()
    for token in Tokenizer(formula).items:
        if token.type == "OPERAND" and token.subtype == "ERROR":
            raise ValuationProjectionError(f"Formula contains an error token: {sheet}!{coordinate} {formula!r}.")
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue
        expanded = _expand_reference(sheet, token.value)
        if expanded is None:
            # Function names, quoted strings, and booleans are different token subtypes.
            # Any remaining RANGE operand is therefore a defined name.  The accepted
            # canonical dependency surface must be positional only at presentation level
            # and uses direct cell references internally.
            raise ValuationProjectionError(
                f"Unresolved defined/range dependency {token.value!r} in {sheet}!{coordinate}."
            )
        for dependency in expanded:
            if dependency[0] not in ALLOWED_IC_DEPENDENCY_SHEETS:
                raise ValuationProjectionError(
                    f"Canonical Investment Case output depends on disallowed sheet "
                    f"{dependency[0]!r}: {sheet}!{coordinate}."
                )
            result.add(dependency)
    return tuple(sorted(result))


def investment_case_dependency_closure(donor_workbook: Path) -> InvestmentCaseDependencyClosure:
    workbook = load_workbook(donor_workbook, data_only=False, read_only=False)
    try:
        missing_sheets = sorted(ALLOWED_IC_DEPENDENCY_SHEETS - set(workbook.sheetnames))
        if missing_sheets:
            raise ValuationProjectionError(f"Source-native donor lacks sheets: {missing_sheets!r}.")
        seeds: set[tuple[str, str]] = {
            _cell_key(sheet.replace("{ticker}", "ANF"), coordinate)
            for sheet, coordinate in canonical_investment_case_defined_names().values()
        }
        minimum_column, minimum_row, maximum_column, maximum_row = range_boundaries(
            CANONICAL_VALUATION_MATRIX_RANGE
        )
        seeds.update(
            (IC_DATA_SHEET, f"{get_column_letter(column)}{row}")
            for row in range(minimum_row, maximum_row + 1)
            for column in range(minimum_column, maximum_column + 1)
        )
        queue = list(sorted(seeds))
        seen: set[tuple[str, str]] = set()
        while queue:
            sheet, coordinate = queue.pop()
            key = (sheet, coordinate)
            if key in seen:
                continue
            seen.add(key)
            value = workbook[sheet][coordinate].value
            if isinstance(value, str) and value.startswith("="):
                for dependency in _formula_dependencies(sheet, coordinate, value):
                    if dependency not in seen:
                        queue.append(dependency)
        formula_count = 0
        value_count = 0
        blank_count = 0
        counts: dict[str, int] = {}
        for sheet, coordinate in seen:
            counts[sheet] = counts.get(sheet, 0) + 1
            value = workbook[sheet][coordinate].value
            if isinstance(value, str) and value.startswith("="):
                formula_count += 1
            elif value is None:
                blank_count += 1
            else:
                value_count += 1
        return InvestmentCaseDependencyClosure(
            seed_count=len(seeds),
            cell_count=len(seen),
            formula_count=formula_count,
            value_count=value_count,
            blank_count=blank_count,
            cells_by_sheet=dict(sorted(counts.items())),
            cells=tuple(sorted(seen)),
        )
    finally:
        workbook.close()


def _number_text(value: int | float | Decimal) -> str:
    parsed = Decimal(str(value))
    normalized = format(parsed, "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    return normalized or "0"


def _value_mutation(
    *,
    sheet: str,
    coordinate: str,
    value: Any,
    style_source_cell: str | None = None,
    number_format_code: str | None = None,
    semantic_owner: str,
) -> FormulaAwareCellMutation:
    if value is None:
        return FormulaAwareCellMutation(
            sheet,
            coordinate,
            "CLEAR_CONTENTS",
            style_source_cell=style_source_cell,
            semantic_owner=semantic_owner,
        )
    if isinstance(value, bool):
        return FormulaAwareCellMutation(
            sheet,
            coordinate,
            "SET_VALUE",
            "1" if value else "0",
            "boolean",
            number_format_code,
            style_source_cell,
            semantic_owner,
        )
    if isinstance(value, (int, float, Decimal)):
        return FormulaAwareCellMutation(
            sheet,
            coordinate,
            "SET_VALUE",
            _number_text(value),
            "number",
            number_format_code,
            style_source_cell,
            semantic_owner,
        )
    if isinstance(value, str) and value.startswith("="):
        return FormulaAwareCellMutation(
            sheet,
            coordinate,
            "SET_FORMULA",
            value[1:],
            None,
            number_format_code,
            style_source_cell,
            semantic_owner,
        )
    if isinstance(value, str):
        return FormulaAwareCellMutation(
            sheet,
            coordinate,
            "SET_VALUE",
            value,
            "text",
            number_format_code,
            style_source_cell,
            semantic_owner,
        )
    raise ValuationProjectionError(
        f"Unsupported donor value type at {sheet}!{coordinate}: {type(value).__name__}."
    )


def _matrix_number_format(coordinate: str) -> str | None:
    column = _CELL_RE.fullmatch(coordinate).group(1)  # type: ignore[union-attr]
    if column in {"BI", "BN"}:
        return "$0.00"
    if column in {"BF", "BJ", "BL", "BM"}:
        return "0.0%"
    if column in {"BE", "BG", "BH"}:
        return "#,##0.0"
    return None


def _references_name(formula: str, name: str) -> bool:
    return re.search(rf"(?<![A-Za-z0-9_.]){re.escape(name)}(?![A-Za-z0-9_.])", formula) is not None


def build_valuation_projection_plan(
    *,
    base_workbook: Path,
    donor_workbook: Path,
    exhaustive_audit_dir: Path,
) -> ValuationProjectionPlan:
    audit_manifest_path = exhaustive_audit_dir / "audit_manifest.json"
    historical = _load_json(exhaustive_audit_dir / "CURRENT_HISTORICAL_RECONCILIATION.json")
    forward = _load_json(exhaustive_audit_dir / "FORWARD_SUMMARY_REQUIREMENTS.json")
    ownership = _load_json(exhaustive_audit_dir / "FORMULA_OWNERSHIP_DECISION.json")
    names_review = _load_json(exhaustive_audit_dir / "NAMED_RANGE_HIDDEN_SUPPORT_REVIEW.json")
    base_sha = _sha256_file(base_workbook)
    if base_sha != BASE_SUMMARY_BS_GOLDEN_SHA256:
        raise ValuationProjectionError(f"Summary/BS golden base changed: {base_sha}.")
    if forward.get("required_input_count") != 20:
        raise ValuationProjectionError("Forward summary audit no longer contains exactly 20 inputs.")
    if ownership.get("counts") != {"KEEP_WORKBOOK_OWNED": 1, "RETIRE_DUPLICATE_ENGINE": 74}:
        raise ValuationProjectionError("Formula ownership audit changed from the accepted 1/74 disposition.")

    base = load_workbook(base_workbook, data_only=False, read_only=False)
    donor = load_workbook(donor_workbook, data_only=False, read_only=False)
    mutations: dict[tuple[str, str], FormulaAwareCellMutation] = {}

    def put(item: FormulaAwareCellMutation) -> None:
        mutations[(item.target_sheet, item.target_cell)] = item

    def set_value(
        coordinate: str,
        value: Any,
        *,
        sheet: str = VALUATION_SHEET,
        style_source_cell: str | None = None,
        number_format_code: str | None = None,
        owner: str,
    ) -> None:
        put(
            _value_mutation(
                sheet=sheet,
                coordinate=coordinate,
                value=value,
                style_source_cell=style_source_cell,
                number_format_code=number_format_code,
                semantic_owner=owner,
            )
        )

    try:
        valuation = base[VALUATION_SHEET]

        # Audit-authoritative historical/current consumer corrections.
        checks = historical["protected_workbook_consumer_checks"]
        liquidity_issues = tuple(
            record
            for record in checks["debt_liquidity_comparisons"]
            if record["classification"] != "EXACT"
        )
        securities_issues = tuple(
            record
            for record in checks["marketable_and_net_cash_comparisons"]
            if record["classification"] != "EXACT"
        )
        if len(liquidity_issues) != 47 or len(securities_issues) != 5:
            raise ValuationProjectionError(
                f"Accepted consumer issue inventory changed: {len(liquidity_issues)}/47 liquidity, "
                f"{len(securities_issues)}/5 securities/net cash."
            )
        for record in (*liquidity_issues, *securities_issues):
            set_value(record["cell"], record["source_native_value"], owner="accepted_source_native_consumer")

        # Invalid P&L interest coverage is retired without a replacement ratio.
        interest_cells = tuple(f"{column}88" for column in "BCDEFGHIJKLM") + ("B149",)
        for coordinate in interest_cells[:-1]:
            set_value(coordinate, None, owner="retired_invalid_legacy_semantic")
        set_value(
            "A88",
            "Interest coverage (P&L TTM) — unavailable",
            owner="presentation_status",
        )
        set_value(
            "B149",
            "Interest coverage unavailable under the accepted definition.",
            owner="presentation_status",
        )

        # Clear every populated cell in both duplicate Valuation forward engines.
        for minimum_column, minimum_row, maximum_column, maximum_row in (
            (15, 48, 27, 75),  # O48:AA75 Thesis Bridge
            (1, 192, 35, 261),  # A192:AI261 detailed engine
        ):
            for row in range(minimum_row, maximum_row + 1):
                for column in range(minimum_column, maximum_column + 1):
                    cell = valuation.cell(row, column)
                    if cell.value is not None:
                        set_value(cell.coordinate, None, owner="retired_duplicate_forward_engine")

        # Concise retirement states preserve the investor capability without a second engine.
        set_value(
            "O48",
            "Forward valuation is owned by Investment Case",
            style_source_cell="O48",
            owner="presentation_status",
        )
        set_value(
            "O49",
            "See the compact canonical summary at A192:F198.",
            style_source_cell="O49",
            owner="presentation_status",
        )

        # Compact canonical forward-summary surface.
        for column in "ABCDEF":
            set_value(
                f"{column}192",
                "Forward Valuation Summary" if column == "A" else None,
                style_source_cell="B192",
                owner="valuation_presentation",
            )
        headers = {
            "A193": ("Metric", "B193"),
            "B193": ("Current", "D193"),
            "C193": ("Bear", "D193"),
            "D193": ("Base", "D193"),
            "E193": ("Bull", "D193"),
            "F193": ("State / context", "F193"),
        }
        for coordinate, (value, style_source) in headers.items():
            set_value(
                coordinate,
                value,
                style_source_cell=style_source,
                owner="valuation_presentation",
            )
        metric_rows = {
            194: ("GAAP diluted EPS ($/share)", "$0.00"),
            195: ("Adjusted EBITDA ($m)", "#,##0.0"),
            196: ("FCF per diluted share ($/share)", "$0.00"),
            197: ("Blended value per share ($/share)", "$0.00"),
            198: ("Upside / downside", "0.0%"),
        }
        scenarios = ("Current", "Bear", "Base", "Bull")
        tokens = {
            194: "GAAP_EPS",
            195: "Adjusted_EBITDA",
            196: "FCF_Per_Share",
            197: "Blended_Value_Per_Share",
            198: "Upside_Downside",
        }
        compact_cells: list[str] = []
        for row, (label, number_format) in metric_rows.items():
            set_value(
                f"A{row}",
                label,
                style_source_cell=f"B{row}",
                owner="valuation_presentation",
            )
            for index, scenario in enumerate(scenarios, start=2):
                coordinate = f"{get_column_letter(index)}{row}"
                compact_cells.append(coordinate)
                set_value(
                    coordinate,
                    f"=IC_{scenario}_{tokens[row]}",
                    style_source_cell=f"D{row}",
                    number_format_code=number_format,
                    owner="canonical_investment_case_link",
                )
            context = (
                "Unavailable — current market price not populated"
                if row == 198
                else "Canonical Investment Case output"
            )
            set_value(
                f"F{row}",
                context,
                style_source_cell=f"F{row}",
                owner="valuation_presentation",
            )
        set_value(
            "A200",
            "Detailed valuation mechanics remain in Investment Case.",
            style_source_cell="F194",
            owner="presentation_status",
        )

        # Bounded investor-facing product corrections.
        label_updates = {
            "A73": "Net debt (core debt less cash; excludes securities)",
            "A77": "Core net cash (cash less core debt)",
            "A78": "Net cash incl. marketable securities",
            "A116": "Market-linked — unavailable (current price not populated)",
            "A124": "No funded core debt instruments as of 2026-Q1",
            "A125": "Leases separate; undrawn ABL remains in liquidity.",
        }
        style_sources = {"A124": "A139", "A125": "B149"}
        for coordinate, value in label_updates.items():
            set_value(
                coordinate,
                value,
                style_source_cell=style_sources.get(coordinate),
                owner="valuation_presentation",
            )

        # Install the exact accepted IC matrix/name outputs plus their direct dependency closure.
        closure = investment_case_dependency_closure(donor_workbook)
        donor_formats = {
            (IC_VISIBLE_SHEET, f"{column}{row}"): format_code
            for row, format_code in ((91, "#,##0.0"), (98, "$0.00"), (99, "$0.00"))
            for column in "BCDE"
        }
        for sheet, coordinate in closure.cells:
            donor_value = donor[sheet][coordinate].value
            base_value = base[sheet][coordinate].value
            if donor_value is None and base_value is None:
                continue
            number_format = donor_formats.get((sheet, coordinate))
            if sheet == IC_DATA_SHEET:
                number_format = _matrix_number_format(coordinate)
            set_value(
                coordinate,
                donor_value,
                sheet=sheet,
                number_format_code=number_format,
                owner="investment_case_source_native_dependency",
            )

        # Retire legacy names bound to the former Valuation engine.  Four aliases
        # remain only to keep unrelated Hidden Value formulas truthfully source-bound.
        valuation_name_records = names_review["valuation_related_defined_names"]
        preserved_names = {
            "CompanyOperatingMargin_Latest",
            "OperatingMargin_Latest",
            "CompanyOperatingMargin_TTM",
            "FCF_TTM_Pos_Years",
            "Pos_FCF_Ratio",
            "Interest_Coverage",
        }
        alias_rebindings = {
            "Price": "'ANF_Investment_Case'!$D$106",
            "FCF_TTM": "'ANF_Investment_Case'!$B$94",
            "Shares": "'ANF_Investment_Case'!$B$95",
            "FCF_Yield": "'Hidden_Value_Audit'!$H$2",
        }
        reviewed_names = {record["name"] for record in valuation_name_records}
        name_deletions = tuple(sorted(reviewed_names - preserved_names - set(alias_rebindings)))

        # Any formula still using a deleted legacy name must either be inside the
        # already-retired Valuation engine or be retired explicitly on the old IC surface.
        for worksheet in base.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    formula = cell.value
                    if not (isinstance(formula, str) and formula.startswith("=")):
                        continue
                    referenced = [name for name in name_deletions if _references_name(formula, name)]
                    if not referenced:
                        continue
                    target = (worksheet.title, cell.coordinate)
                    if target in mutations:
                        continue
                    if worksheet.title == IC_VISIBLE_SHEET:
                        set_value(
                            cell.coordinate,
                            None,
                            sheet=worksheet.title,
                            owner="retired_legacy_investment_case_alias_consumer",
                        )
                        continue
                    raise ValuationProjectionError(
                        f"Deleting legacy names would break unrelated formula {worksheet.title}!{cell.coordinate}: "
                        f"{referenced!r}."
                    )

        defined_names: list[DefinedNameMutation] = [
            DefinedNameMutation(name, "DELETE") for name in name_deletions
        ]
        defined_names.extend(
            DefinedNameMutation(name, "UPSERT", reference)
            for name, reference in sorted(alias_rebindings.items())
        )
        for name, (sheet, coordinate) in sorted(canonical_investment_case_defined_names().items()):
            resolved_sheet = sheet.replace("{ticker}", "ANF")
            column = "".join(character for character in coordinate if character.isalpha())
            row = "".join(character for character in coordinate if character.isdigit())
            defined_names.append(
                DefinedNameMutation(name, "UPSERT", f"'{resolved_sheet}'!${column}${row}")
            )

        merge_mutations = (
            WorksheetMergeMutation(VALUATION_SHEET, "B192:S192", "DELETE"),
            WorksheetMergeMutation(VALUATION_SHEET, "B193:C193", "DELETE"),
            WorksheetMergeMutation(VALUATION_SHEET, "B194:C194", "DELETE"),
            WorksheetMergeMutation(VALUATION_SHEET, "B195:C195", "DELETE"),
            WorksheetMergeMutation(VALUATION_SHEET, "B196:C196", "DELETE"),
            WorksheetMergeMutation(VALUATION_SHEET, "B197:C197", "DELETE"),
            WorksheetMergeMutation(VALUATION_SHEET, "B198:C198", "DELETE"),
            WorksheetMergeMutation(VALUATION_SHEET, "A192:F192", "ADD"),
        )
        row_mutations = tuple(
            WorksheetRowMutation(VALUATION_SHEET, row, True) for row in range(201, 262)
        )

        retire_cells = tuple(ownership["retire_duplicate_engine_cells"])
        if len(retire_cells) != 74:
            raise ValuationProjectionError("Accepted retirement ledger no longer contains 74 formulas.")
        for coordinate in retire_cells:
            mutation = mutations.get((VALUATION_SHEET, coordinate))
            if mutation is None or mutation.mode != "CLEAR_CONTENTS":
                raise ValuationProjectionError(f"Retired formula is not cleared: Valuation!{coordinate}.")

        # The only legitimate Valuation-local formula is deliberately untouched.
        if valuation["AI139"].value != '=IFERROR(MATCH(1,\'Hidden_Value_Flags\'!$L$2:$L$100,0)+1,"")':
            raise ValuationProjectionError("Valuation!AI139 changed from the accepted formula.")
        if (VALUATION_SHEET, "AI139") in mutations:
            raise ValuationProjectionError("Projection may not mutate Valuation!AI139.")

        ordered_mutations = tuple(
            sorted(
                mutations.values(),
                key=lambda item: (
                    item.target_sheet,
                    int(_CELL_RE.fullmatch(item.target_cell).group(2)),  # type: ignore[union-attr]
                    _CELL_RE.fullmatch(item.target_cell).group(1),  # type: ignore[union-attr]
                ),
            )
        )
        ordered_names = tuple(sorted(defined_names, key=lambda item: item.name.casefold()))
        formula_plan = [
            {
                "cell": f"{item.target_sheet}!{item.target_cell}",
                "formula": item.value,
                "owner": item.semantic_owner,
            }
            for item in ordered_mutations
            if item.mode == "SET_FORMULA"
        ]
        defined_name_plan = [asdict(item) for item in ordered_names]
        digest_payload = {
            "cell_mutations": [asdict(item) for item in ordered_mutations],
            "contract": PROJECTION_CONTRACT,
            "defined_name_mutations": defined_name_plan,
            "ic_dependency_cells": list(closure.cells),
            "merge_mutations": [asdict(item) for item in merge_mutations],
            "row_mutations": [asdict(item) for item in row_mutations],
        }
        return ValuationProjectionPlan(
            contract=PROJECTION_CONTRACT,
            audit_manifest_sha256=_sha256_file(audit_manifest_path),
            base_workbook_sha256=base_sha,
            donor_workbook_sha256=_sha256_file(donor_workbook),
            cell_mutations=ordered_mutations,
            defined_name_mutations=ordered_names,
            merge_mutations=merge_mutations,
            row_mutations=row_mutations,
            ic_dependency_closure=closure,
            liquidity_issue_cells=tuple(record["cell"] for record in liquidity_issues),
            securities_net_cash_issue_cells=tuple(record["cell"] for record in securities_issues),
            interest_coverage_cells=interest_cells,
            old_formula_retirement_cells=retire_cells,
            compact_link_cells=tuple(compact_cells),
            legacy_name_deletions=name_deletions,
            legacy_name_rebindings=dict(sorted(alias_rebindings.items())),
            projection_digest=_digest(digest_payload),
            formula_plan_digest=_digest(formula_plan),
            defined_name_plan_digest=_digest(defined_name_plan),
        )
    finally:
        base.close()
        donor.close()


_PLAN_KEYS = {
    "audit_manifest_sha256",
    "base_workbook_sha256",
    "calculation_metadata_policy_id",
    "cell_mutation_count",
    "cell_mutations",
    "compact_link_cells",
    "contract",
    "defined_name_mutation_count",
    "defined_name_mutations",
    "defined_name_plan_digest",
    "donor_workbook_sha256",
    "formula_plan_digest",
    "ic_dependency_closure",
    "interest_coverage_cells",
    "legacy_name_deletions",
    "legacy_name_rebindings",
    "liquidity_issue_cells",
    "merge_mutations",
    "old_formula_retirement_cells",
    "projection_digest",
    "row_mutations",
    "securities_net_cash_issue_cells",
}
_CELL_MUTATION_KEYS = {
    "mode",
    "number_format_code",
    "semantic_owner",
    "style_source_cell",
    "target_cell",
    "target_sheet",
    "value",
    "value_kind",
}
_DEFINED_NAME_MUTATION_KEYS = {"attr_text", "mode", "name"}
_MERGE_MUTATION_KEYS = {"mode", "range_ref", "target_sheet"}
_ROW_MUTATION_KEYS = {"hidden", "row", "target_sheet"}
_CLOSURE_KEYS = {
    "blank_count",
    "cell_count",
    "cells",
    "cells_by_sheet",
    "formula_count",
    "seed_count",
    "value_count",
}


def _closed_rows(
    value: Any,
    *,
    keys: set[str],
    label: str,
) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        raise ValuationProjectionError(f"{label} must be a list.")
    result: list[dict[str, Any]] = []
    for index, row in enumerate(value):
        if not isinstance(row, dict) or set(row) != keys:
            raise ValuationProjectionError(f"{label}[{index}] is not a closed record.")
        result.append(row)
    return result


def _require_sha256(value: Any, *, label: str) -> str:
    text = str(value or "").casefold()
    if re.fullmatch(r"[0-9a-f]{64}", text) is None:
        raise ValuationProjectionError(f"{label} is not a SHA-256 identity.")
    return text


def load_valuation_projection_plan(
    path: Path | str,
    *,
    expected_projection_digest: str | None = None,
    expected_formula_plan_digest: str | None = None,
    expected_defined_name_plan_digest: str | None = None,
) -> ValuationProjectionPlan:
    """Load one immutable committed Valuation plan and verify all content identities."""

    payload = _load_json(Path(path))
    if set(payload) != _PLAN_KEYS:
        raise ValuationProjectionError("Valuation projection fixture is not a closed schema.")
    if payload["contract"] != PROJECTION_CONTRACT:
        raise ValuationProjectionError("Valuation projection contract changed.")
    if payload["calculation_metadata_policy_id"] != CALCULATION_METADATA_POLICY_ID:
        raise ValuationProjectionError("Valuation calculation-metadata policy changed.")

    cell_rows = _closed_rows(
        payload["cell_mutations"], keys=_CELL_MUTATION_KEYS, label="cell_mutations"
    )
    name_rows = _closed_rows(
        payload["defined_name_mutations"],
        keys=_DEFINED_NAME_MUTATION_KEYS,
        label="defined_name_mutations",
    )
    merge_rows = _closed_rows(
        payload["merge_mutations"], keys=_MERGE_MUTATION_KEYS, label="merge_mutations"
    )
    row_rows = _closed_rows(
        payload["row_mutations"], keys=_ROW_MUTATION_KEYS, label="row_mutations"
    )
    if payload["cell_mutation_count"] != len(cell_rows):
        raise ValuationProjectionError("Valuation cell mutation count mismatch.")
    if payload["defined_name_mutation_count"] != len(name_rows):
        raise ValuationProjectionError("Valuation defined-name mutation count mismatch.")

    closure_payload = payload["ic_dependency_closure"]
    if not isinstance(closure_payload, dict) or set(closure_payload) != _CLOSURE_KEYS:
        raise ValuationProjectionError("Valuation dependency closure is not a closed record.")
    raw_cells = closure_payload["cells"]
    if not isinstance(raw_cells, list) or any(
        not isinstance(item, list) or len(item) != 2 for item in raw_cells
    ):
        raise ValuationProjectionError("Valuation dependency closure cells are malformed.")
    closure = InvestmentCaseDependencyClosure(
        seed_count=int(closure_payload["seed_count"]),
        cell_count=int(closure_payload["cell_count"]),
        formula_count=int(closure_payload["formula_count"]),
        value_count=int(closure_payload["value_count"]),
        blank_count=int(closure_payload["blank_count"]),
        cells_by_sheet={
            str(key): int(value)
            for key, value in dict(closure_payload["cells_by_sheet"]).items()
        },
        cells=tuple((str(item[0]), str(item[1])) for item in raw_cells),
    )
    if closure.cell_count != len(closure.cells):
        raise ValuationProjectionError("Valuation dependency closure count mismatch.")

    cell_mutations = tuple(FormulaAwareCellMutation(**row) for row in cell_rows)
    defined_name_mutations = tuple(DefinedNameMutation(**row) for row in name_rows)
    merge_mutations = tuple(WorksheetMergeMutation(**row) for row in merge_rows)
    row_mutations = tuple(WorksheetRowMutation(**row) for row in row_rows)
    formula_plan = [
        {
            "cell": f"{item.target_sheet}!{item.target_cell}",
            "formula": item.value,
            "owner": item.semantic_owner,
        }
        for item in cell_mutations
        if item.mode == "SET_FORMULA"
    ]
    defined_name_plan = [asdict(item) for item in defined_name_mutations]
    projection_payload = {
        "cell_mutations": [asdict(item) for item in cell_mutations],
        "contract": PROJECTION_CONTRACT,
        "defined_name_mutations": defined_name_plan,
        "ic_dependency_cells": list(closure.cells),
        "merge_mutations": [asdict(item) for item in merge_mutations],
        "row_mutations": [asdict(item) for item in row_mutations],
    }
    actual_projection = _digest(projection_payload)
    actual_formula = _digest(formula_plan)
    actual_names = _digest(defined_name_plan)
    declared_projection = _require_sha256(payload["projection_digest"], label="projection_digest")
    declared_formula = _require_sha256(payload["formula_plan_digest"], label="formula_plan_digest")
    declared_names = _require_sha256(
        payload["defined_name_plan_digest"], label="defined_name_plan_digest"
    )
    if (actual_projection, actual_formula, actual_names) != (
        declared_projection,
        declared_formula,
        declared_names,
    ):
        raise ValuationProjectionError("Valuation projection fixture digest mismatch.")
    expected = (
        expected_projection_digest or declared_projection,
        expected_formula_plan_digest or declared_formula,
        expected_defined_name_plan_digest or declared_names,
    )
    if (declared_projection, declared_formula, declared_names) != expected:
        raise ValuationProjectionError("Valuation projection fixture is not the expected accepted plan.")

    return ValuationProjectionPlan(
        contract=PROJECTION_CONTRACT,
        audit_manifest_sha256=_require_sha256(
            payload["audit_manifest_sha256"], label="audit_manifest_sha256"
        ),
        base_workbook_sha256=_require_sha256(
            payload["base_workbook_sha256"], label="base_workbook_sha256"
        ),
        donor_workbook_sha256=_require_sha256(
            payload["donor_workbook_sha256"], label="donor_workbook_sha256"
        ),
        cell_mutations=cell_mutations,
        defined_name_mutations=defined_name_mutations,
        merge_mutations=merge_mutations,
        row_mutations=row_mutations,
        ic_dependency_closure=closure,
        liquidity_issue_cells=tuple(str(item) for item in payload["liquidity_issue_cells"]),
        securities_net_cash_issue_cells=tuple(
            str(item) for item in payload["securities_net_cash_issue_cells"]
        ),
        interest_coverage_cells=tuple(str(item) for item in payload["interest_coverage_cells"]),
        old_formula_retirement_cells=tuple(
            str(item) for item in payload["old_formula_retirement_cells"]
        ),
        compact_link_cells=tuple(str(item) for item in payload["compact_link_cells"]),
        legacy_name_deletions=tuple(str(item) for item in payload["legacy_name_deletions"]),
        legacy_name_rebindings={
            str(key): str(value) for key, value in dict(payload["legacy_name_rebindings"]).items()
        },
        projection_digest=declared_projection,
        formula_plan_digest=declared_formula,
        defined_name_plan_digest=declared_names,
    )


def write_valuation_projection_plan(
    plan: ValuationProjectionPlan,
    output_path: Path | str,
) -> Path:
    """Write the immutable Valuation projection fixture deterministically."""

    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(_canonical_bytes(plan.as_dict(include_mutations=True)))
    return target


__all__ = [
    "AUDIT_SCHEMA",
    "BASE_SUMMARY_BS_GOLDEN_SHA256",
    "CALCULATION_METADATA_POLICY_ID",
    "PROJECTION_CONTRACT",
    "PROTECTED_ANF_SHA256",
    "VALUATION_CALCULATION_METADATA_POLICY",
    "InvestmentCaseDependencyClosure",
    "ValuationProjectionError",
    "ValuationProjectionPlan",
    "build_valuation_projection_plan",
    "investment_case_dependency_closure",
    "load_valuation_projection_plan",
    "write_valuation_projection_plan",
]
