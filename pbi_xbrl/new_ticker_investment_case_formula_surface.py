"""Compact, fail-closed workbook formulas for the Investment Case product."""
from __future__ import annotations

from typing import Any, Iterable


INVESTMENT_CASE_SHEET = "{ticker}_Investment_Case"
INVESTMENT_CASE_DATA_SHEET = "{ticker}_Investment_Case_Data"
CANONICAL_VALUATION_MATRIX_RANGE = "BB1:BQ25"
CANONICAL_VALUATION_MATRIX_HEADERS = (
    "scenario_id",
    "method_id",
    "metric_id",
    "metric_value",
    "target_value",
    "enterprise_value",
    "equity_value",
    "value_per_share",
    "upside_downside",
    "method_state",
    "entered_weight",
    "effective_weight",
    "weighted_value_per_share",
    "period_id",
    "formula_owner",
    "lineage_state",
)
CANONICAL_SCENARIOS = (
    ("Current baseline", "Current", "B"),
    ("Bear", "Bear", "C"),
    ("Base", "Base", "D"),
    ("Bull", "Bull", "E"),
)
CANONICAL_VALUATION_METHODS = (
    ("pe", "PE", "gaap_diluted_eps", 0),
    ("ev_adjusted_ebitda", "EV_Adjusted_EBITDA", "adjusted_ebitda", 1),
    ("ev_revenue", "EV_Revenue", "revenue", 2),
    ("fcf_yield", "FCF_Yield", "free_cash_flow", 3),
    ("dcf", "DCF", "fcff", 4),
    ("blended", "Blended", "blended_value_per_share", 5),
)
CANONICAL_VALUATION_MATRIX_COLUMNS = (
    "BB",
    "BC",
    "BD",
    "BE",
    "BF",
    "BG",
    "BH",
    "BI",
    "BJ",
    "BK",
    "BL",
    "BM",
    "BN",
    "BO",
    "BP",
    "BQ",
)


def canonical_valuation_matrix_row(scenario_token: str, method_id: str) -> int:
    scenario_index = next(
        index for index, (_label, token, _column) in enumerate(CANONICAL_SCENARIOS) if token == scenario_token
    )
    method_index = next(
        index for index, (candidate, _name_token, _metric_id, _offset) in enumerate(CANONICAL_VALUATION_METHODS)
        if candidate == method_id
    )
    return 2 + scenario_index * len(CANONICAL_VALUATION_METHODS) + method_index


def canonical_valuation_matrix_lookup_expression(
    *,
    support_sheet_reference: str,
    output_column: str,
    scenario_expression: str,
    method_id: str,
) -> str:
    """Resolve exactly one canonical valuation row by scenario and method identity."""

    if output_column not in CANONICAL_VALUATION_MATRIX_COLUMNS:
        raise ValueError(f"Unknown canonical valuation matrix column {output_column!r}.")
    valid_method_ids = {row[0] for row in CANONICAL_VALUATION_METHODS}
    if method_id not in valid_method_ids:
        raise ValueError(f"Unknown canonical valuation method {method_id!r}.")
    method_literal = method_id.replace('"', '""')
    scenario_range = f"{support_sheet_reference}!$BB$2:$BB$25"
    method_range = f"{support_sheet_reference}!$BC$2:$BC$25"
    output_range = f"{support_sheet_reference}!${output_column}$2:${output_column}$25"
    match_expression = (
        f'MATCH(1,INDEX(({scenario_range}={scenario_expression})*'
        f'({method_range}="{method_literal}"),0),0)'
    )
    return (
        f'IF(COUNTIFS({scenario_range},{scenario_expression},'
        f'{method_range},"{method_literal}")<>1,"",'
        f'INDEX({output_range},{match_expression}))'
    )


def canonical_investment_case_defined_names() -> dict[str, tuple[str, str]]:
    """Return stable Investment Case output names and their owned destinations."""

    result: dict[str, tuple[str, str]] = {}
    operating_rows = {
        "GAAP_EPS": 98,
        "Adjusted_EBITDA": 91,
        "FCF_Per_Share": 99,
    }
    for _label, scenario_token, scenario_column in CANONICAL_SCENARIOS:
        for output_token, row in operating_rows.items():
            result[f"IC_{scenario_token}_{output_token}"] = (
                INVESTMENT_CASE_SHEET,
                f"{scenario_column}{row}",
            )
        for method_id, name_token, _metric_id, _offset in CANONICAL_VALUATION_METHODS[:-1]:
            result[f"IC_{scenario_token}_{name_token}_Value_Per_Share"] = (
                INVESTMENT_CASE_DATA_SHEET,
                f"BI{canonical_valuation_matrix_row(scenario_token, method_id)}",
            )
        blended_row = canonical_valuation_matrix_row(scenario_token, "blended")
        result[f"IC_{scenario_token}_Blended_Value_Per_Share"] = (
            INVESTMENT_CASE_DATA_SHEET,
            f"BI{blended_row}",
        )
        result[f"IC_{scenario_token}_Upside_Downside"] = (
            INVESTMENT_CASE_DATA_SHEET,
            f"BJ{blended_row}",
        )
    return result


def _safe_expression(
    operands: Iterable[str],
    expression: str,
    *,
    invalid_conditions: Iterable[str] = (),
) -> str:
    """Return nested Excel guards with arithmetic only in the innermost branch."""

    result = expression
    for condition in reversed(tuple(invalid_conditions)):
        result = f'IF({condition},"",{result})'
    for operand in reversed(tuple(operands)):
        result = f'IF(NOT(ISNUMBER({operand})),"",{result})'
    return result


def _safe_formula(
    operands: Iterable[str],
    expression: str,
    *,
    invalid_conditions: Iterable[str] = (),
) -> str:
    return f"={_safe_expression(operands, expression, invalid_conditions=invalid_conditions)}"


def _text_literal(value: str) -> str:
    return f'="{value.replace(chr(34), chr(34) * 2)}"'


def apply_investment_case_formula_surface(
    workbook: Any,
    enabled_formula_ids: set[str],
) -> None:
    """Build the final read-through Investment Case workflow."""

    from pbi_xbrl import standard_template_formula_contract as contract

    if INVESTMENT_CASE_SHEET not in workbook.sheetnames:
        return
    ws = workbook[INVESTMENT_CASE_SHEET]

    contract._prepare_investment_case_scenario_layout(ws)
    contract._clear_range(ws, "A13:Q240")
    for range_ref in contract.user_input_surface_targets("{ticker}_Investment_Case"):
        contract._clear_range(ws, range_ref)

    section_titles = {
        13: "Model Data and Guidance",
        38: "Scenario Assumptions",
        71: "Selected Scenario Incremental Bridge",
        82: "Scenario Output Comparison",
        103: "Valuation and DCF Assumptions",
        119: "Valuation Summary",
        129: "What the Market Is Pricing",
        141: "Guidance-Implied Earnings",
        152: "DCF and Equity Value",
        167: "Calculation Details",
        190: "Sensitivity Tables",
        217: "Key Debates and Invalidators",
    }
    for row, title in section_titles.items():
        ws.cell(row, 1).value = title

    ws["A14"] = (
        "Read-only model data separates the latest completed year, exact four-quarter TTM, "
        "company guidance and the value passed into the scenario baseline."
    )
    for column, label in enumerate(
        (
            "Metric",
            "",
            "",
            "",
            "",
            "Active value",
            "Active source",
        ),
        start=1,
    ):
        ws.cell(15, column).value = label
    for row, label in (
        (16, "Market data"),
        (20, "Operating performance"),
        (30, "Cash flow and capital allocation"),
    ):
        ws[f"A{row}"] = label

    ws["A39"] = (
        "Current baseline is read-only. Bear, Base and Bull are explicit editable cases; "
        "a blank scenario input inherits the current baseline while numeric zero remains valid. "
        "Enter amounts in the units shown in the Metric label."
    )
    ws["A40"] = "Revenue Build"
    for column, label in enumerate(
        ("Driver", "Current baseline", "Bear", "Base", "Bull", "Source / state"),
        start=1,
    ):
        ws.cell(41, column).value = label
    ws["A42"] = "Revenue scenario mode"
    ws["A46"] = "Brand"
    ws["A49"] = "Geography"
    ws["A53"] = "Brand and Geography are alternative views and must not be added together."
    ws["A54"] = "Operating Drivers"
    ws["A61"] = "Cash Flow and Capital Allocation"
    for header_row in (55, 62):
        for column, label in enumerate(
            ("Driver", "Current baseline", "Bear", "Base", "Bull", "Source / state"),
            start=1,
        ):
            ws.cell(header_row, column).value = label
    ws["A69"] = "Selected scenario"
    ws["C69"] = "Blank selection uses Base"

    for column, label in enumerate(
        (
            "Metric / effect",
            "Current baseline",
            "Selected scenario",
            "Incremental change",
            "Resulting output",
        ),
        start=1,
    ):
        ws.cell(72, column).value = label
    ws["A81"] = (
        "The bridge shows total selected-scenario changes. Effects are not presented as "
        "independent contributions when drivers interact."
    )

    for column, label in enumerate(
        ("Output", "Current baseline", "Bear", "Base", "Bull"),
        start=1,
    ):
        ws.cell(83, column).value = label
    ws["A84"] = "Operating Results"
    ws["A93"] = "Cash Flow and Capital Structure"
    ws["A97"] = "Per-Share and Value"

    ws["A104"] = (
        "Blank assumptions remain unavailable. Manual inputs feed the existing canonical "
        "valuation and DCF formulas; no fallback multiple, tax rate or price is supplied."
    )
    for column, label in enumerate(
        ("Assumption", "Current / default", "Manual input", "Active assumption", "State", "Explanation"),
        start=1,
    ):
        ws.cell(105, column).value = label
    ws["A115"] = "Method weights (%)"
    for column, label in zip("BCDEF", ("P/E", "EV / Adj. EBITDA", "EV / Revenue", "FCF yield", "DCF"), strict=True):
        ws[f"{column}116"] = label

    for column, label in enumerate(
        (
            "Method",
            "Metric used",
            "Target",
            "Equity value ($m)",
            "Value/share ($/share)",
            "Upside/downside (%)",
            "State",
            "Weight (%)",
        ),
        start=1,
    ):
        ws.cell(120, column).value = label

    for column, label in enumerate(("Metric", "Value", "Interpretation"), start=1):
        ws.cell(130, column).value = label
    for column, label in enumerate(
        ("Metric", "Company guidance", "Period", "Model translation", "Selected scenario", "State"),
        start=1,
    ):
        ws.cell(142, column).value = label

    ws["A153"] = (
        "DCF estimates enterprise value from future FCFF. Forecast cash flows and terminal value "
        "are discounted using WACC; net debt/cash then bridges enterprise value to equity value."
    )
    ws["A154"] = "DCF Forecast"
    ws["G154"] = "DCF Valuation Bridge"
    for column, label in enumerate(("FCFF component", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"), start=1):
        ws.cell(155, column).value = label
    ws["G155"] = "Metric"
    ws["H155"] = "Value"
    ws["I155"] = "State"

    ws["A168"] = (
        "These are forecast/model calculations; historical source-native buyback detail belongs "
        "to the later Capital Return pass."
    )
    ws["A169"] = "Operating, EPS and cash-flow chain"
    for column, label in enumerate(("Calculation", "Basis", "Result"), start=1):
        ws.cell(170, column).value = label
    ws["A180"] = "Buyback and share-count chain"
    for column, label in enumerate(("Calculation", "Basis", "Result"), start=1):
        ws.cell(181, column).value = label

    ws["A191"] = (
        "Each family activates only when its required assumptions and positive axis steps "
        "are numeric. Blank setup cells never become zero."
    )
    for row, title in (
        (192, "P/E x EPS sensitivity"),
        (198, "EV / adjusted EBITDA sensitivity"),
        (204, "FCF-yield sensitivity"),
        (210, "DCF WACC x terminal-growth sensitivity"),
    ):
        ws[f"A{row}"] = title

    for column, label in enumerate(
        ("Debate", "Risk or invalidator", "", "", "", "", "State", "Watch next", "", "", "", "Source"),
        start=1,
    ):
        ws.cell(218, column).value = label

    if "investment_case_scenario_formulas" not in enabled_formula_ids:
        return

    set_formula = contract._set_formula
    support_num = contract._ic_support_numeric_expression
    support_text = contract._ic_support_text_expression
    number_format = contract._ic_number_format

    fy_period = support_text("market_input|revenue", "AK")
    ttm_period = support_text("market_input|revenue", "AL")
    fy_guidance_period = support_text("market_input|revenue_growth", "AM")
    quarter_guidance_period = support_text("market_input|revenue_growth", "AN")
    set_formula(
        ws["B15"],
        f'=IF({fy_period}="","Model default (year)","Model default ("&LEFT({fy_period},4)&" year)")',
        "General",
    )
    set_formula(ws["C15"], '="Model default (TTM)"', "General")
    set_formula(
        ws["D15"],
        f'=IF({fy_guidance_period}="","Guidance (year)","Guidance ("&RIGHT({fy_guidance_period},4)&" year)")',
        "General",
    )
    set_formula(
        ws["E15"],
        f'=IF({quarter_guidance_period}="","Guidance (quarter)","Guidance ("&{quarter_guidance_period}&")")',
        "General",
    )

    model_rows = {
        "price": 17,
        "diluted_shares": 18,
        "net_debt": 19,
        "revenue": 21,
        "revenue_growth": 22,
        "gross_margin": 23,
        "operating_margin": 24,
        "base_ebitda": 25,
        "base_ebitda_margin": 26,
        "adjusted_ebitda": 27,
        "adjusted_ebitda_margin": 28,
        "net_income": 29,
        "free_cash_flow": 31,
        "depreciation_amortization": 32,
        "capital_expenditures": 33,
        "working_capital_investment": 34,
        "buyback_cash": 35,
    }
    metric_units = {
        "price": "$/share",
        "diluted_shares": "m shares",
        "net_debt": "$m",
        "revenue": "$m",
        "revenue_growth": "%",
        "gross_margin": "%",
        "operating_margin": "%",
        "base_ebitda": "$m",
        "base_ebitda_margin": "%",
        "adjusted_ebitda": "$m",
        "adjusted_ebitda_margin": "%",
        "net_income": "$m",
        "free_cash_flow": "$m",
        "depreciation_amortization": "$m",
        "capital_expenditures": "$m",
        "working_capital_investment": "$m",
        "buyback_cash": "$m",
        "buyback_execution_price": "$/share",
        "share_issuance": "m shares",
        "tax_rate": "%",
        "target_pe": "x",
        "target_ev_adjusted_ebitda": "x",
        "target_ev_revenue": "x",
        "target_fcf_yield": "%",
        "dcf_revenue_growth": "%",
        "dcf_wacc": "%",
        "dcf_terminal_growth": "%",
        "dcf_forecast_years": "years",
    }
    for metric_id, row in model_rows.items():
        slot = f"market_input|{metric_id}"
        fmt = number_format(metric_units[metric_id])
        set_formula(ws[f"A{row}"], f"={support_text(slot, 'Y')}", "General")
        if metric_id == "buyback_cash":
            set_formula(ws[f"B{row}"], '=""', fmt)
            set_formula(ws[f"C{row}"], '=""', fmt)
        else:
            set_formula(ws[f"B{row}"], f"={support_num(slot, 'AA')}", fmt)
            set_formula(ws[f"C{row}"], f"={support_num(slot, 'AB')}", fmt)
        set_formula(ws[f"D{row}"], f"={support_text(slot, 'AC')}", "General")
        set_formula(ws[f"E{row}"], f"={support_text(slot, 'AF')}", "General")
        resolved = support_num(slot, "AI")
        basis_kind = support_text(slot, "AW")
        row_fy_period = support_text(slot, "AK")
        set_formula(ws[f"F{row}"], f'=IF(ISNUMBER({resolved}),{resolved},"")', fmt)
        set_formula(
            ws[f"G{row}"],
            (
                f'=IF(ISNUMBER(F{row}),'
                f'IF({basis_kind}="latest_completed_fy",'
                f'IF({row_fy_period}="","Model default (year)",'
                f'"Model default ("&LEFT({row_fy_period},4)&" year)"),'
                f'"Model default (TTM)"),'
                f'IF(OR({basis_kind}="manual_only",{basis_kind}="historical_context_only"),'
                f'"Manual input required","Unavailable"))'
            ),
            "General",
        )

    dimension_mode = 'IF($B$42="","Total Company",$B$42)'
    total_slot = "segment_input|001"
    set_formula(ws["B44"], f"={support_num(total_slot, 'AI')}", "#,##0.0;-#,##0.0")
    ws["A44"] = "Total Company revenue ($m)"
    ws["A45"] = "Total Company revenue growth (%)"
    set_formula(ws["B45"], f'=IF(ISNUMBER($F${model_rows["revenue_growth"]}),$F${model_rows["revenue_growth"]},"")', "0.0%;-0.0%")
    set_formula(ws["F44"], '="Read-only Total Company tie-out"', "General")
    set_formula(
        ws["F45"],
        '=IF(ISNUMBER(B45),"Used only in Total Company mode","Manual input required")',
        "General",
    )
    for column in "CDE":
        ws[f"{column}45"].number_format = "0.0%;-0.0%"

    segment_rows = ((47, 2), (48, 3), (50, 4), (51, 5), (52, 6))
    for row, slot_index in segment_rows:
        slot = f"segment_input|{slot_index:03d}"
        set_formula(ws[f"A{row}"], f"={support_text(slot, 'Y')}", "General")
        set_formula(ws[f"B{row}"], f"={support_num(slot, 'AI')}", "#,##0.0;-#,##0.0")
        basis_kind = support_text(slot, "AW")
        basis_period = support_text(slot, "AX")
        set_formula(
            ws[f"F{row}"],
            (
                f'=IF(NOT(ISNUMBER(B{row})),"Unavailable",'
                f'IF({basis_kind}="exact_four_quarter_ttm",'
                f'IF({basis_period}="","Model default (TTM)",{basis_period}),'
                f'IF({basis_kind}="latest_completed_fy",'
                f'IF({basis_period}="","Model default (year)",'
                f'"Model default ("&LEFT({basis_period},4)&" year)"),"Resolved")))'
            ),
            "General",
        )
        for column in "CDE":
            ws[f"{column}{row}"].number_format = "0.0%;-0.0%"
    set_formula(
        ws["A46"],
        (
            f'=IF({support_text("segment_input|002", "AV")}="","Brand",'
            f'{support_text("segment_input|002", "AV")})'
        ),
        "General",
    )
    set_formula(
        ws["A49"],
        (
            f'=IF({support_text("segment_input|004", "AV")}="","Geography",'
            f'{support_text("segment_input|004", "AV")})'
        ),
        "General",
    )

    driver_specs = (
        (56, "Gross margin (%)", "gross_margin"),
        (57, "Operating margin (%)", "operating_margin"),
        (58, "Base EBITDA margin (%)", "base_ebitda_margin"),
        (59, "Adjusted EBITDA margin (%)", "adjusted_ebitda_margin"),
        (60, "Tax rate (%)", "tax_rate"),
        (63, "Capital expenditure ($m)", "capital_expenditures"),
        (64, "Working-capital investment ($m)", "working_capital_investment"),
        (65, "Buyback cash ($m)", "buyback_cash"),
        (66, "Buyback execution price ($/share)", "buyback_execution_price"),
        (67, "Share issuance (m)", "share_issuance"),
        (68, "Net cash / debt ($m)", "net_debt"),
    )
    for row, label, metric_id in driver_specs:
        ws[f"A{row}"] = label
        slot = f"market_input|{metric_id}"
        fmt = number_format(metric_units[metric_id])
        source_row = model_rows.get(metric_id)
        if source_row is None:
            set_formula(ws[f"B{row}"], f"={support_num(slot, 'AI')}", fmt)
            basis_kind = support_text(slot, "AW")
            set_formula(
                ws[f"F{row}"],
                (
                    f'=IF(ISNUMBER(B{row}),"Model default (TTM)",'
                    f'IF(OR({basis_kind}="manual_only",{basis_kind}="historical_context_only"),'
                    f'"Manual input required","Unavailable"))'
                ),
                "General",
            )
        else:
            set_formula(ws[f"B{row}"], f'=IF(ISNUMBER($F${source_row}),$F${source_row},"")', fmt)
            set_formula(
                ws[f"F{row}"],
                f'=IF(ISNUMBER(B{row}),$G${source_row},"Manual input required")',
                "General",
            )
        for column in "CDE":
            ws[f"{column}{row}"].number_format = fmt

    selected_scenario = 'IF($B$69="","Base",$B$69)'

    def driver_value(row: int, scenario_column: str) -> str:
        if scenario_column == "B":
            return f"$B${row}"
        return (
            f'IF(${scenario_column}${row}="",$B${row},'
            f'IF(ISNUMBER(${scenario_column}${row}),${scenario_column}${row},"Unavailable"))'
        )

    def selected_driver(row: int) -> str:
        selected_override = (
            f'INDEX($C${row}:$E${row},1,'
            f'MATCH({selected_scenario},$C$41:$E$41,0))'
        )
        return (
            f'IF({selected_override}="",$B${row},'
            f'IF(ISNUMBER({selected_override}),{selected_override},"Unavailable"))'
        )

    segment_group_rows = {47: "$A$46", 48: "$A$46", 50: "$A$49", 51: "$A$49", 52: "$A$49"}

    def segment_impact(scenario_column: str) -> str:
        if scenario_column == "B":
            return "0"
        terms = []
        for row, group_cell in segment_group_rows.items():
            terms.append(
                (
                    f'IF(LOWER({dimension_mode})=LOWER({group_cell}),'
                    f'IF(ISNUMBER($B${row}),'
                    f'IF(ISNUMBER(${scenario_column}${row}),$B${row}*${scenario_column}${row},0),0),0)'
                )
            )
        return f'IF(LOWER({dimension_mode})="total company",0,{"+".join(terms)})'

    def effective_revenue_growth(scenario_column: str) -> str:
        if scenario_column == "B":
            return "$B$45"
        return (
            f'IF(LOWER({dimension_mode})="total company",'
            f'{driver_value(45, scenario_column)},$B$45)'
        )

    output_rows = {
        "revenue": 85,
        "revenue_growth": 86,
        "gross_margin": 87,
        "operating_margin": 88,
        "operating_income": 89,
        "base_ebitda": 90,
        "adjusted_ebitda": 91,
        "net_income": 92,
        "free_cash_flow": 94,
        "diluted_shares": 95,
        "net_debt": 96,
        "eps": 98,
        "fcf_per_share": 99,
        "selected_value": 100,
        "selected_upside": 101,
    }
    output_labels = {
        85: "Revenue ($m)",
        86: "Revenue growth (%)",
        87: "Gross margin (%)",
        88: "Operating margin (%)",
        89: "Operating income ($m)",
        90: "Base EBITDA ($m)",
        91: "Adjusted EBITDA ($m)",
        92: "GAAP net income ($m)",
        94: "FCF ($m)",
        95: "Diluted shares (m)",
        96: "Net cash / debt ($m)",
        98: "GAAP diluted EPS ($/share)",
        99: "FCF/share ($/share)",
        100: "Selected value/share ($/share)",
        101: "Selected upside/downside (%)",
    }
    for row, label in output_labels.items():
        ws[f"A{row}"] = label

    base_revenue = f"$F${model_rows['revenue']}"
    base_growth = f"$F${model_rows['revenue_growth']}"
    base_op_margin = f"$F${model_rows['operating_margin']}"
    base_net_income = f"$F${model_rows['net_income']}"
    base_shares = f"$F${model_rows['diluted_shares']}"
    base_fcf = f"$F${model_rows['free_cash_flow']}"
    base_capex = f"$F${model_rows['capital_expenditures']}"
    scenario_columns = "BCDE"
    for output_column in scenario_columns:
        revenue_growth = effective_revenue_growth(output_column)
        gross_margin = driver_value(56, output_column)
        operating_margin = driver_value(57, output_column)
        base_ebitda_margin = driver_value(58, output_column)
        adjusted_ebitda_margin = driver_value(59, output_column)
        tax_rate = driver_value(60, output_column)
        capex = driver_value(63, output_column)
        working_capital = driver_value(64, output_column)
        buyback_cash = driver_value(65, output_column)
        execution_price = driver_value(66, output_column)
        issuance = driver_value(67, output_column)
        net_debt = driver_value(68, output_column)
        impact = segment_impact(output_column)

        if output_column == "B":
            set_formula(ws[f"{output_column}85"], f'=IF(ISNUMBER({base_revenue}),{base_revenue},"")', "#,##0.0;-#,##0.0")
        else:
            set_formula(
                ws[f"{output_column}85"],
                _safe_formula(
                    (base_revenue, base_growth, revenue_growth),
                    f"{base_revenue}*(1+{revenue_growth}-{base_growth})+{impact}",
                ),
                "#,##0.0;-#,##0.0",
            )
        if output_column == "B":
            set_formula(
                ws["B86"],
                f'=IF(ISNUMBER({base_growth}),{base_growth},"")',
                "0.0%;-0.0%",
            )
        else:
            prior_revenue = f"{base_revenue}/(1+{base_growth})"
            set_formula(
                ws[f"{output_column}86"],
                _safe_formula(
                    (f"{output_column}85", base_revenue, base_growth),
                    f"{output_column}85/({prior_revenue})-1",
                    invalid_conditions=(
                        f"{base_revenue}=0",
                        f"1+{base_growth}=0",
                    ),
                ),
                "0.0%;-0.0%",
            )
        set_formula(ws[f"{output_column}87"], f'=IF(ISNUMBER({gross_margin}),{gross_margin},"")', "0.0%;-0.0%")
        set_formula(ws[f"{output_column}88"], f'=IF(ISNUMBER({operating_margin}),{operating_margin},"")', "0.0%;-0.0%")
        set_formula(
            ws[f"{output_column}89"],
            _safe_formula((f"{output_column}85", f"{output_column}88"), f"{output_column}85*{output_column}88"),
            "#,##0.0;-#,##0.0",
        )
        set_formula(
            ws[f"{output_column}90"],
            _safe_formula((f"{output_column}85", base_ebitda_margin), f"{output_column}85*{base_ebitda_margin}"),
            "#,##0.0;-#,##0.0",
        )
        set_formula(
            ws[f"{output_column}91"],
            _safe_formula((f"{output_column}85", adjusted_ebitda_margin), f"{output_column}85*{adjusted_ebitda_margin}"),
            "#,##0.0;-#,##0.0",
        )
        baseline_operating_income = f"{base_revenue}*{base_op_margin}"
        if output_column == "B":
            set_formula(ws["B92"], f'=IF(ISNUMBER({base_net_income}),{base_net_income},"")', "#,##0.0;-#,##0.0")
            set_formula(ws["B94"], f'=IF(ISNUMBER({base_fcf}),{base_fcf},"")', "#,##0.0;-#,##0.0")
            set_formula(ws["B95"], f'=IF(ISNUMBER({base_shares}),{base_shares},"")', "#,##0.000;-#,##0.000")
        else:
            net_income_formula = (
                f'=IF(NOT(ISNUMBER({output_column}89)),"",'
                f'IF(NOT(ISNUMBER({base_revenue})),"",'
                f'IF(NOT(ISNUMBER({base_op_margin})),"",'
                f'IF(NOT(ISNUMBER({base_net_income})),"",'
                f'IF(ABS({output_column}89-{baseline_operating_income})<0.0000001,{base_net_income},'
                f'IF(NOT(ISNUMBER({tax_rate})),"",'
                f'IF({tax_rate}<0,"",'
                f'IF({tax_rate}>1,"",{base_net_income}+'
                f'({output_column}89-{baseline_operating_income})*(1-{tax_rate})))))))))'
            )
            set_formula(ws[f"{output_column}92"], net_income_formula, "#,##0.0;-#,##0.0")
            fcf_formula = (
                f'=IF(NOT(ISNUMBER({base_fcf})),"",'
                f'IF(NOT(ISNUMBER({output_column}89)),"",'
                f'IF(NOT(ISNUMBER({base_revenue})),"",'
                f'IF(NOT(ISNUMBER({base_op_margin})),"",'
                f'IF(ABS({output_column}89-{baseline_operating_income})<0.0000001,{base_fcf},'
                f'IF(NOT(ISNUMBER({tax_rate})),"",'
                f'IF({tax_rate}<0,"",'
                f'IF({tax_rate}>1,"",'
                f'IF(NOT(ISNUMBER({capex})),"",'
                f'IF(NOT(ISNUMBER({base_capex})),"",'
                f'IF(NOT(ISNUMBER({working_capital})),"",'
                f'{base_fcf}+({output_column}89-{baseline_operating_income})*(1-{tax_rate})'
                f'-({capex}-{base_capex})-{working_capital})))))))))))'
            )
            set_formula(ws[f"{output_column}94"], fcf_formula, "#,##0.0;-#,##0.0")
            set_formula(
                ws[f"{output_column}95"],
                "="
                + contract._ic_safe_resulting_shares_formula(
                    base_shares,
                    buyback_cash,
                    execution_price,
                    issuance,
                ),
                "#,##0.000;-#,##0.000",
            )
        set_formula(ws[f"{output_column}96"], f'=IF(ISNUMBER({net_debt}),{net_debt},"")', "#,##0.0;-#,##0.0")
        set_formula(
            ws[f"{output_column}98"],
            _safe_formula(
                (f"{output_column}92", f"{output_column}95"),
                f"{output_column}92/{output_column}95",
                invalid_conditions=(f"{output_column}95<=0",),
            ),
            "$0.00;-$0.00",
        )
        set_formula(
            ws[f"{output_column}99"],
            _safe_formula(
                (f"{output_column}94", f"{output_column}95"),
                f"{output_column}94/{output_column}95",
                invalid_conditions=(f"{output_column}95<=0",),
            ),
            "$0.00;-$0.00",
        )

    def selected_output(row: int) -> str:
        return f'INDEX($C${row}:$E${row},1,MATCH({selected_scenario},$C$83:$E$83,0))'

    assumption_specs = (
        (
            106,
            "Current share price ($/share)",
            "price",
            "Market price used to calculate market capitalization and upside/downside.",
        ),
        (
            107,
            "Target P/E (x)",
            "target_pe",
            "How much equity value is assigned for each dollar of GAAP EPS.",
        ),
        (
            108,
            "Target EV / Adjusted EBITDA (x)",
            "target_ev_adjusted_ebitda",
            "Enterprise value relative to adjusted EBITDA, before the effect of net debt.",
        ),
        (
            109,
            "Target EV / Revenue (x)",
            "target_ev_revenue",
            "Enterprise value relative to sales; useful when comparing different margin levels.",
        ),
        (
            110,
            "Target FCF yield (%)",
            "target_fcf_yield",
            "Free cash flow divided by equity value. A higher target yield implies a lower valuation.",
        ),
        (
            111,
            "DCF revenue growth (%)",
            "dcf_revenue_growth",
            "Annual revenue growth assumed during the explicit DCF forecast.",
        ),
        (
            112,
            "WACC (%)",
            "dcf_wacc",
            "Discount rate applied to future FCFF. A higher WACC generally lowers value.",
        ),
        (
            113,
            "Terminal growth (%)",
            "dcf_terminal_growth",
            "Long-term growth assumed after the explicit forecast. It must remain below WACC.",
        ),
        (
            114,
            "Forecast period (years)",
            "dcf_forecast_years",
            "Number of explicit forecast years before the terminal value.",
        ),
    )
    assumption_rows: dict[str, int] = {}
    for row, label, metric_id, explanation in assumption_specs:
        assumption_rows[metric_id] = row
        ws[f"A{row}"] = label
        slot = f"market_input|{metric_id}"
        fmt = number_format(metric_units[metric_id])
        set_formula(ws[f"B{row}"], f"={support_num(slot, 'AI')}", fmt)
        set_formula(ws[f"D{row}"], f'=IF(ISNUMBER(C{row}),C{row},IF(ISNUMBER(B{row}),B{row},""))', fmt)
        set_formula(
            ws[f"E{row}"],
            (
                f'=IF(ISNUMBER(C{row}),"Manual input",'
                f'IF(ISNUMBER(B{row}),"Model default","Manual input required"))'
            ),
            "General",
        )
        ws[f"F{row}"] = explanation
        ws[f"C{row}"].number_format = fmt

    support_sheet = workbook[INVESTMENT_CASE_DATA_SHEET]
    support_name = contract.quote_sheetname(INVESTMENT_CASE_DATA_SHEET)
    visible_name = contract.quote_sheetname(INVESTMENT_CASE_SHEET)

    def visible_ref(coordinate: str) -> str:
        column = "".join(character for character in coordinate if character.isalpha())
        row = "".join(character for character in coordinate if character.isdigit())
        return f"{visible_name}!${column}${row}"

    def matrix_ref(column: str, scenario_token: str, method_id: str) -> str:
        return f"{support_name}!${column}${canonical_valuation_matrix_row(scenario_token, method_id)}"

    def selected_matrix(column: str, method_id: str) -> str:
        return canonical_valuation_matrix_lookup_expression(
            support_sheet_reference=support_name,
            output_column=column,
            scenario_expression=selected_scenario,
            method_id=method_id,
        )

    def dcf_fcff_expression(scenario_column: str, year_index: int) -> str:
        revenue = visible_ref(f"{scenario_column}85")
        margin = visible_ref(f"{scenario_column}88")
        tax_rate = visible_ref(f"{scenario_column}60")
        capex = visible_ref(f"{scenario_column}63")
        working_capital = visible_ref(f"{scenario_column}64")
        growth = visible_ref(f"D{assumption_rows['dcf_revenue_growth']}")
        base_revenue_ref = visible_ref(f"F{model_rows['revenue']}")
        dna = visible_ref(f"F{model_rows['depreciation_amortization']}")
        forecast_revenue = f"{revenue}*(1+{growth})^{year_index}"
        return (
            f"({forecast_revenue}*{margin}*(1-{tax_rate})+"
            f"{dna}/{base_revenue_ref}*{forecast_revenue}-"
            f"{capex}/{base_revenue_ref}*{forecast_revenue}-{working_capital})"
        )

    def dcf_enterprise_value_formula(scenario_column: str) -> str:
        revenue = visible_ref(f"{scenario_column}85")
        margin = visible_ref(f"{scenario_column}88")
        tax_rate = visible_ref(f"{scenario_column}60")
        capex = visible_ref(f"{scenario_column}63")
        working_capital = visible_ref(f"{scenario_column}64")
        growth = visible_ref(f"D{assumption_rows['dcf_revenue_growth']}")
        wacc = visible_ref(f"D{assumption_rows['dcf_wacc']}")
        terminal_growth = visible_ref(f"D{assumption_rows['dcf_terminal_growth']}")
        years = visible_ref(f"D{assumption_rows['dcf_forecast_years']}")
        base_revenue_ref = visible_ref(f"F{model_rows['revenue']}")
        dna = visible_ref(f"F{model_rows['depreciation_amortization']}")
        fcff = {year: dcf_fcff_expression(scenario_column, year) for year in range(1, 6)}
        present_values = "+".join(
            f"IF({years}>={year},{fcff[year]}/(1+{wacc})^{year},0)"
            for year in range(1, 6)
        )
        terminal_fcff = fcff[5]
        for year in range(4, 0, -1):
            terminal_fcff = f"IF({years}={year},{fcff[year]},{terminal_fcff})"
        expression = (
            f"({present_values})+({terminal_fcff})*(1+{terminal_growth})/"
            f"({wacc}-{terminal_growth})/(1+{wacc})^{years}"
        )
        return _safe_formula(
            (
                years,
                revenue,
                growth,
                margin,
                tax_rate,
                dna,
                base_revenue_ref,
                capex,
                working_capital,
                wacc,
                terminal_growth,
            ),
            expression,
            invalid_conditions=(
                f"{years}<1",
                f"{years}>5",
                f"MOD({years},1)<>0",
                f"{base_revenue_ref}=0",
                f"{tax_rate}<0",
                f"{tax_rate}>1",
                f"{wacc}<={terminal_growth}",
                f"{wacc}<=-1",
            ),
        )

    if "investment_case_valuation_matrix_formulas" in enabled_formula_ids:
        contract._clear_range(support_sheet, CANONICAL_VALUATION_MATRIX_RANGE)
        contract._remove_data_validations_overlapping(support_sheet, (CANONICAL_VALUATION_MATRIX_RANGE,))
        for column_index, header in enumerate(CANONICAL_VALUATION_MATRIX_HEADERS, start=54):
            support_sheet.cell(1, column_index).value = header

        period_formula = (
            '=IFERROR(INDEX($AL$2:$AL$201,MATCH("market_input|revenue",$W$2:$W$201,0)),"")'
        )
        target_rows = {
            "pe": assumption_rows["target_pe"],
            "ev_adjusted_ebitda": assumption_rows["target_ev_adjusted_ebitda"],
            "ev_revenue": assumption_rows["target_ev_revenue"],
            "fcf_yield": assumption_rows["target_fcf_yield"],
            "dcf": assumption_rows["dcf_wacc"],
        }
        metric_rows = {
            "pe": output_rows["eps"],
            "ev_adjusted_ebitda": output_rows["adjusted_ebitda"],
            "ev_revenue": output_rows["revenue"],
            "fcf_yield": output_rows["free_cash_flow"],
        }
        weight_columns = {
            "pe": "B",
            "ev_adjusted_ebitda": "C",
            "ev_revenue": "D",
            "fcf_yield": "E",
            "dcf": "F",
        }
        method_weight_range = f"{visible_name}!$B$117:$F$117"
        invalid_method_weights = (
            f"OR(COUNT({method_weight_range})<>COUNTA({method_weight_range}),"
            f'COUNTIF({method_weight_range},"<0")>0,'
            f'COUNTIF({method_weight_range},">1")>0)'
        )
        for scenario_label, scenario_token, scenario_column in CANONICAL_SCENARIOS:
            first_method_row = canonical_valuation_matrix_row(scenario_token, "pe")
            last_method_row = canonical_valuation_matrix_row(scenario_token, "dcf")
            net_debt = visible_ref(f"{scenario_column}{output_rows['net_debt']}")
            shares = visible_ref(f"{scenario_column}{output_rows['diluted_shares']}")
            price = visible_ref(f"D{assumption_rows['price']}")
            for method_id, _name_token, metric_id, _offset in CANONICAL_VALUATION_METHODS:
                row = canonical_valuation_matrix_row(scenario_token, method_id)
                support_sheet[f"BB{row}"] = scenario_label
                support_sheet[f"BC{row}"] = method_id
                support_sheet[f"BD{row}"] = metric_id
                support_sheet[f"BP{row}"] = "investment_case_scenario_valuation"
                support_sheet[f"BQ{row}"] = "canonical_formula"
                set_formula(support_sheet[f"BO{row}"], period_formula, "General")
                if method_id == "blended":
                    for column in ("BE", "BF", "BG", "BH"):
                        set_formula(support_sheet[f"{column}{row}"], '=""', "General")
                    denominator = f"SUM(BM{first_method_row}:BM{last_method_row})"
                    numerator = f"SUM(BN{first_method_row}:BN{last_method_row})"
                    set_formula(
                        support_sheet[f"BI{row}"],
                        (
                            f'=IF({invalid_method_weights},"",'
                            f'IF({denominator}<=0,"",'
                            f'IF(ABS({denominator}-1)>0.0000001,"",{numerator})))'
                        ),
                        "$0.00;-$0.00",
                    )
                    set_formula(
                        support_sheet[f"BJ{row}"],
                        _safe_formula((f"BI{row}", price), f"BI{row}/{price}-1", invalid_conditions=(f"{price}<=0",)),
                        "0.0%;-0.0%",
                    )
                    set_formula(
                        support_sheet[f"BK{row}"],
                        (
                            f'=IF({invalid_method_weights},"Invalid method weight",'
                            f'IF(ISNUMBER(BI{row}),"Available methods weighted to 100%",'
                            f'IF({denominator}<=0,"Unavailable",'
                            f'"Available-method weights must sum to 100%")))'
                        ),
                        "General",
                    )
                    set_formula(support_sheet[f"BL{row}"], '=""', "0.0%;-0.0%")
                    set_formula(
                        support_sheet[f"BM{row}"],
                        f'=IF({invalid_method_weights},"",{denominator})',
                        "0.0%;-0.0%",
                    )
                    set_formula(
                        support_sheet[f"BN{row}"],
                        f'=IF({invalid_method_weights},"",IF(ISNUMBER(BI{row}),BI{row},""))',
                        "$0.00;-$0.00",
                    )
                    continue

                if method_id == "dcf":
                    metric = dcf_fcff_expression(scenario_column, 1)
                    target = visible_ref(f"D{target_rows[method_id]}")
                    dcf_base_revenue = visible_ref(f"F{model_rows['revenue']}")
                    dcf_tax_rate = visible_ref(f"{scenario_column}60")
                    set_formula(
                        support_sheet[f"BE{row}"],
                        _safe_formula(
                            (
                                visible_ref(f"{scenario_column}85"),
                                visible_ref(f"{scenario_column}88"),
                                visible_ref(f"{scenario_column}60"),
                                visible_ref(f"{scenario_column}63"),
                                visible_ref(f"{scenario_column}64"),
                                visible_ref(f"D{assumption_rows['dcf_revenue_growth']}"),
                                visible_ref(f"F{model_rows['revenue']}"),
                                visible_ref(f"F{model_rows['depreciation_amortization']}"),
                            ),
                            metric,
                            invalid_conditions=(
                                f"{dcf_base_revenue}=0",
                                f"{dcf_tax_rate}<0",
                                f"{dcf_tax_rate}>1",
                            ),
                        ),
                        "#,##0.0;-#,##0.0",
                    )
                    set_formula(support_sheet[f"BF{row}"], f'=IF(ISNUMBER({target}),{target},"")', "0.0%;-0.0%")
                    set_formula(support_sheet[f"BG{row}"], dcf_enterprise_value_formula(scenario_column), "#,##0.0;-#,##0.0")
                    set_formula(
                        support_sheet[f"BH{row}"],
                        _safe_formula((f"BG{row}", net_debt), f"BG{row}-{net_debt}"),
                        "#,##0.0;-#,##0.0",
                    )
                    set_formula(
                        support_sheet[f"BI{row}"],
                        _safe_formula((f"BH{row}", shares), f"BH{row}/{shares}", invalid_conditions=(f"{shares}<=0",)),
                        "$0.00;-$0.00",
                    )
                else:
                    metric = visible_ref(f"{scenario_column}{metric_rows[method_id]}")
                    target = visible_ref(f"D{target_rows[method_id]}")
                    set_formula(support_sheet[f"BE{row}"], f'=IF(ISNUMBER({metric}),{metric},"")', contract._ic_number_format({"pe": "$/share", "ev_adjusted_ebitda": "$m", "ev_revenue": "$m", "fcf_yield": "$m"}[method_id]))
                    set_formula(support_sheet[f"BF{row}"], f'=IF(ISNUMBER({target}),{target},"")', "0.0%;-0.0%" if method_id == "fcf_yield" else "0.00x;-0.00x")
                    if method_id == "pe":
                        equity_formula = _safe_formula((metric, target, shares), f"{metric}*{target}*{shares}", invalid_conditions=(f"{target}<=0", f"{shares}<=0"))
                        value_formula = _safe_formula((metric, target), f"{metric}*{target}", invalid_conditions=(f"{target}<=0",))
                        set_formula(support_sheet[f"BH{row}"], equity_formula, "#,##0.0;-#,##0.0")
                        set_formula(support_sheet[f"BG{row}"], _safe_formula((f"BH{row}", net_debt), f"BH{row}+{net_debt}"), "#,##0.0;-#,##0.0")
                        set_formula(support_sheet[f"BI{row}"], value_formula, "$0.00;-$0.00")
                    elif method_id in {"ev_adjusted_ebitda", "ev_revenue"}:
                        set_formula(support_sheet[f"BG{row}"], _safe_formula((metric, target), f"{metric}*{target}", invalid_conditions=(f"{target}<=0",)), "#,##0.0;-#,##0.0")
                        set_formula(support_sheet[f"BH{row}"], _safe_formula((f"BG{row}", net_debt), f"BG{row}-{net_debt}"), "#,##0.0;-#,##0.0")
                        set_formula(support_sheet[f"BI{row}"], _safe_formula((f"BH{row}", shares), f"BH{row}/{shares}", invalid_conditions=(f"{shares}<=0",)), "$0.00;-$0.00")
                    else:
                        set_formula(support_sheet[f"BH{row}"], _safe_formula((metric, target), f"{metric}/{target}", invalid_conditions=(f"{target}<=0",)), "#,##0.0;-#,##0.0")
                        set_formula(support_sheet[f"BG{row}"], _safe_formula((f"BH{row}", net_debt), f"BH{row}+{net_debt}"), "#,##0.0;-#,##0.0")
                        set_formula(support_sheet[f"BI{row}"], _safe_formula((f"BH{row}", shares), f"BH{row}/{shares}", invalid_conditions=(f"{shares}<=0",)), "$0.00;-$0.00")

                set_formula(
                    support_sheet[f"BJ{row}"],
                    _safe_formula((f"BI{row}", price), f"BI{row}/{price}-1", invalid_conditions=(f"{price}<=0",)),
                    "0.0%;-0.0%",
                )
                set_formula(support_sheet[f"BK{row}"], f'=IF(ISNUMBER(BI{row}),"Available","Unavailable")', "General")
                weight = visible_ref(f"{weight_columns[method_id]}117")
                set_formula(support_sheet[f"BL{row}"], f'=IF({weight}="","",{weight})', "0.0%;-0.0%")
                set_formula(
                    support_sheet[f"BM{row}"],
                    (
                        f'=IF({invalid_method_weights},"",'
                        f'IF(NOT(ISNUMBER(BI{row})),0,IF(BL{row}="",0,BL{row})))'
                    ),
                    "0.0%;-0.0%",
                )
                set_formula(
                    support_sheet[f"BN{row}"],
                    (
                        f'=IF({invalid_method_weights},"",'
                        f'IF(NOT(ISNUMBER(BM{row})),"",'
                        f'IF(BM{row}=0,0,IF(NOT(ISNUMBER(BI{row})),"",BI{row}*BM{row}))))'
                    ),
                    "$0.00;-$0.00",
                )

    ws["G115"] = (
        "Method weights (%) - Blended value/share, enter percentages that sum to 100% "
        "across available methods; blank or 0 excludes a method."
    )
    for column in "BCDEF":
        ws[f"{column}117"].number_format = "0.0%;-0.0%"

    selected_price = f"$D${assumption_rows['price']}"
    selected_shares = selected_output(output_rows["diluted_shares"])
    selected_net_debt = selected_output(output_rows["net_debt"])
    valuation_methods = (
        (121, "P/E", "pe", "$/share", "0.00x;-0.00x"),
        (122, "EV / adjusted EBITDA", "ev_adjusted_ebitda", "$m", "0.00x;-0.00x"),
        (123, "EV / revenue", "ev_revenue", "$m", "0.00x;-0.00x"),
        (124, "FCF yield", "fcf_yield", "$m", "0.0%;-0.0%"),
        (125, "DCF", "dcf", "$m", "0.0%;-0.0%"),
    )
    for row, method, method_id, metric_unit, target_format in valuation_methods:
        ws[f"A{row}"] = method
        set_formula(ws[f"B{row}"], f'={selected_matrix("BE", method_id)}', number_format(metric_unit))
        set_formula(ws[f"C{row}"], f'={selected_matrix("BF", method_id)}', target_format)
        set_formula(ws[f"D{row}"], f'={selected_matrix("BH", method_id)}', "#,##0.0;-#,##0.0")
        set_formula(ws[f"E{row}"], f'={selected_matrix("BI", method_id)}', "$0.00;-$0.00")
        set_formula(ws[f"F{row}"], f'={selected_matrix("BJ", method_id)}', "0.0%;-0.0%")
        set_formula(ws[f"G{row}"], f'={selected_matrix("BK", method_id)}', "General")
        set_formula(ws[f"H{row}"], f'={selected_matrix("BL", method_id)}', "0.0%;-0.0%")

    ws["A126"] = "Weighted value/share ($/share)"
    set_formula(ws["E126"], f'={selected_matrix("BI", "blended")}', "$0.00;-$0.00")
    set_formula(ws["F126"], f'={selected_matrix("BJ", "blended")}', "0.0%;-0.0%")
    set_formula(ws["G126"], f'={selected_matrix("BK", "blended")}', "General")
    ws["A127"] = "Valuation method discipline"
    set_formula(
        ws["G127"],
        '=IF(COUNTIF(G121:G125,"Available")=0,"No method has all required inputs",'
        'COUNTIF(G121:G125,"Available")&" method(s) available")',
        "General",
    )
    set_formula(
        ws["H127"],
        '="Unavailable methods are excluded regardless of entered weight."',
        "General",
    )

    for (_scenario_label, scenario_token, output_column) in CANONICAL_SCENARIOS:
        set_formula(
            ws[f"{output_column}100"],
            f'={matrix_ref("BI", scenario_token, "blended")}',
            "$0.00;-$0.00",
        )
        set_formula(
            ws[f"{output_column}101"],
            f'={matrix_ref("BJ", scenario_token, "blended")}',
            "0.0%;-0.0%",
        )

    bridge_specs = (
        (73, "Revenue growth (%)", "$B$86", selected_output(86), selected_output(85), "%", "$m"),
        (74, "Operating margin (%)", "$B$57", selected_driver(57), selected_output(89), "%", "$m"),
        (75, "Capital expenditure ($m)", "$B$63", selected_driver(63), selected_output(94), "$m", "$m"),
        (76, "Diluted shares (m)", "$B$95", selected_output(95), selected_output(95), "m shares", "m shares"),
        (77, "Revenue ($m)", "$B$85", selected_output(85), selected_output(85), "$m", "$m"),
        (78, "Operating income ($m)", "$B$89", selected_output(89), selected_output(89), "$m", "$m"),
        (79, "FCF ($m)", "$B$94", selected_output(94), selected_output(94), "$m", "$m"),
        (80, "GAAP diluted EPS ($/share)", "$B$98", selected_output(98), selected_output(98), "$/share", "$/share"),
    )
    for row, label, baseline, selected, result, delta_unit, result_unit in bridge_specs:
        ws[f"A{row}"] = label
        set_formula(ws[f"B{row}"], f'=IF(ISNUMBER({baseline}),{baseline},"")', number_format(delta_unit))
        set_formula(ws[f"C{row}"], f'=IF(ISNUMBER({selected}),{selected},"")', number_format(delta_unit))
        set_formula(
            ws[f"D{row}"],
            _safe_formula((baseline, selected), f"{selected}-{baseline}"),
            number_format(delta_unit),
        )
        set_formula(ws[f"E{row}"], f'=IF(ISNUMBER({result}),{result},"")', number_format(result_unit))

    market_rows_out = (
        (131, "Current share price ($/share)", selected_price, "Accepted price or manual input", "$/share"),
        (132, "Scenario diluted shares (m)", selected_shares, "Selected scenario share count", "m shares"),
        (133, "Market capitalization ($m)", _safe_expression((selected_price, selected_shares), f"{selected_price}*{selected_shares}"), "Price x diluted shares", "$m"),
        (134, "Scenario net cash / debt ($m)", selected_net_debt, "Selected scenario net debt", "$m"),
        (135, "Enterprise value ($m)", _safe_expression(("B133", "B134"), "B133+B134"), "Market cap plus net debt", "$m"),
        (
            136,
            "Implied P/E (x)",
            _safe_expression((selected_price, selected_output(98)), f"{selected_price}/{selected_output(98)}", invalid_conditions=(f"{selected_output(98)}<=0",)),
            "Current price / selected EPS",
            "x",
        ),
        (
            137,
            "Implied EV / revenue (x)",
            _safe_expression(("B135", selected_output(85)), f"B135/{selected_output(85)}", invalid_conditions=(f"{selected_output(85)}<=0",)),
            "Market EV / selected revenue",
            "x",
        ),
        (
            138,
            "Implied EV / adjusted EBITDA (x)",
            _safe_expression(("B135", selected_output(91)), f"B135/{selected_output(91)}", invalid_conditions=(f"{selected_output(91)}<=0",)),
            "Market EV / selected adjusted EBITDA",
            "x",
        ),
        (139, "Implied terminal growth (%)", "", "Market EV, selected FCF and WACC", "%"),
    )
    for row, label, expression, interpretation, unit in market_rows_out:
        ws[f"A{row}"] = label
        if row == 139:
            scenario_fcf = selected_output(94)
            wacc = f"$D${assumption_rows['dcf_wacc']}"
            expression = _safe_expression(
                ("B135", scenario_fcf, wacc),
                f"(B135*{wacc}-{scenario_fcf})/(B135+{scenario_fcf})",
                invalid_conditions=("B135<=0", f"{scenario_fcf}<=0", f"{wacc}<=0", f"B135+{scenario_fcf}=0"),
            )
        set_formula(ws[f"B{row}"], f"={expression}", number_format(unit))
        set_formula(
            ws[f"C{row}"],
            f'=IF(ISNUMBER(B{row}),"{interpretation}","Unavailable | {interpretation}")',
            "General",
        )

    guidance_specs = (
        (143, "FY revenue growth (%)", "revenue_growth", "AC", "AM", "AY", "AD", "AE", 86, "%"),
        (144, "FY implied revenue ($m)", "revenue_growth", "AC", "AM", "AY", "AD", "AE", 85, "$m"),
        (145, "FY operating margin (%)", "operating_margin", "AC", "AM", "AY", "AD", "AE", 88, "%"),
        (146, "FY implied operating income ($m)", "operating_margin", "AC", "AM", "AY", "AD", "AE", 89, "$m"),
        (147, "FY adjusted earnings/share ($/share)", "adjusted_eps_guidance", "AC", "AM", "AY", "AD", "AE", 98, "$/share"),
        (148, "Latest-quarter revenue growth (%)", "revenue_growth", "AF", "AN", "AZ", "AG", "AH", 0, "%"),
        (149, "Latest-quarter operating margin (%)", "operating_margin", "AF", "AN", "AZ", "AG", "AH", 0, "%"),
        (150, "Latest-quarter adjusted EPS ($/share)", "adjusted_eps_guidance", "AF", "AN", "AZ", "AG", "AH", 0, "$/share"),
    )
    for row, label, metric_id, display_col, period_col, state_col, low_col, high_col, output_row, unit in guidance_specs:
        slot = f"market_input|{metric_id}"
        ws[f"A{row}"] = label
        numeric_state = support_text(slot, state_col)
        low = support_num(slot, low_col)
        high = support_num(slot, high_col)
        midpoint = _safe_expression((low, high), f"({low}+{high})/2")
        set_formula(ws[f"B{row}"], f"={support_text(slot, display_col)}", "General")
        set_formula(ws[f"C{row}"], f"={support_text(slot, period_col)}", "General")
        translation = midpoint
        if row == 146:
            translation = _safe_expression(("D144", low, high), f"D144*({low}+{high})/2")
        elif row == 144:
            translation = _safe_expression(
                (f"$F${model_rows['revenue']}", low, high),
                f"$F${model_rows['revenue']}*(1+({low}+{high})/2)",
            )
        set_formula(ws[f"D{row}"], f"={translation}", number_format(unit))
        if output_row:
            set_formula(ws[f"E{row}"], f'=IF(ISNUMBER({selected_output(output_row)}),{selected_output(output_row)},"")', number_format(unit))
        else:
            set_formula(ws[f"E{row}"], '="Quarterly comparison only"', "General")
        set_formula(
            ws[f"F{row}"],
            (
                f'=IF({numeric_state}="typed_range","Typed range",'
                f'IF({numeric_state}="typed_point","Typed point",'
                f'IF({numeric_state}="typed_approximate_point","Approximate guidance",'
                f'IF({numeric_state}="typed_minimum_point","Minimum guidance",'
                f'IF({numeric_state}="qualitative_only","Qualitative guidance","Unavailable")))))'
            ),
            "General",
        )

    dcf_labels = (
        "Revenue ($m)",
        "Operating margin (%)",
        "EBIT ($m)",
        "Taxes ($m)",
        "D&A ($m)",
        "Capital expenditure ($m)",
        "Working-capital investment ($m)",
        "FCFF ($m)",
        "Discount factor (x)",
        "Present value ($m)",
    )
    for row, label in enumerate(dcf_labels, start=156):
        ws[f"A{row}"] = label
    dcf_growth = f"$D${assumption_rows['dcf_revenue_growth']}"
    dcf_wacc = f"$D${assumption_rows['dcf_wacc']}"
    dcf_terminal = f"$D${assumption_rows['dcf_terminal_growth']}"
    dcf_years = f"$D${assumption_rows['dcf_forecast_years']}"
    terminal_fcff = f"INDEX($B$163:$F$163,1,{dcf_years})"
    terminal_discount = f"INDEX($B$164:$F$164,1,{dcf_years})"
    valid_horizon_conditions = (f"{dcf_years}<1", f"{dcf_years}>5", f"MOD({dcf_years},1)<>0")
    selected_revenue = selected_output(85)
    selected_margin = selected_output(88)
    selected_tax = selected_driver(60)
    selected_dna = f"$F${model_rows['depreciation_amortization']}"
    selected_capex = selected_driver(63)
    selected_wc = selected_driver(64)
    for year_index, column in enumerate("BCDEF", start=1):
        previous_revenue = selected_revenue if year_index == 1 else f"{chr(ord(column) - 1)}156"
        set_formula(
            ws[f"{column}156"],
            _safe_formula((dcf_years, previous_revenue, dcf_growth), f"{previous_revenue}*(1+{dcf_growth})", invalid_conditions=(f"{dcf_years}<{year_index}",)),
            "#,##0.0;-#,##0.0",
        )
        set_formula(ws[f"{column}157"], _safe_formula((f"{column}156", selected_margin), selected_margin), "0.0%;-0.0%")
        set_formula(ws[f"{column}158"], _safe_formula((f"{column}156", f"{column}157"), f"{column}156*{column}157"), "#,##0.0;-#,##0.0")
        set_formula(ws[f"{column}159"], _safe_formula((f"{column}158", selected_tax), f"-{column}158*{selected_tax}", invalid_conditions=(f"{selected_tax}<0", f"{selected_tax}>1")), "#,##0.0;-#,##0.0")
        set_formula(ws[f"{column}160"], _safe_formula((f"{column}156", selected_dna, base_revenue), f"{selected_dna}/{base_revenue}*{column}156", invalid_conditions=(f"{base_revenue}=0",)), "#,##0.0;-#,##0.0")
        set_formula(ws[f"{column}161"], _safe_formula((f"{column}156", selected_capex, base_revenue), f"-{selected_capex}/{base_revenue}*{column}156", invalid_conditions=(f"{base_revenue}=0",)), "#,##0.0;-#,##0.0")
        set_formula(ws[f"{column}162"], _safe_formula((f"{column}156", selected_wc), f"-{selected_wc}"), "#,##0.0;-#,##0.0")
        set_formula(ws[f"{column}163"], _safe_formula((f"{column}158", f"{column}159", f"{column}160", f"{column}161", f"{column}162"), f"{column}158+{column}159+{column}160+{column}161+{column}162"), "#,##0.0;-#,##0.0")
        set_formula(ws[f"{column}164"], _safe_formula((f"{column}163", dcf_wacc), f"1/(1+{dcf_wacc})^{year_index}", invalid_conditions=(f"{dcf_wacc}<=-1",)), "0.000x")
        set_formula(ws[f"{column}165"], _safe_formula((f"{column}163", f"{column}164"), f"{column}163*{column}164"), "#,##0.0;-#,##0.0")

    dcf_summary = (
        (156, "Forecast period", f'=IF(ISNUMBER({dcf_years}),{dcf_years},"")', "0"),
        (157, "WACC", f'=IF(ISNUMBER({dcf_wacc}),{dcf_wacc},"")', "0.0%;-0.0%"),
        (158, "Terminal growth", f'=IF(ISNUMBER({dcf_terminal}),{dcf_terminal},"")', "0.0%;-0.0%"),
        (159, "Present value of forecast FCFF", _safe_formula((dcf_years,), "SUM(B165:F165)", invalid_conditions=(*valid_horizon_conditions, f"COUNT(B165:F165)<>{dcf_years}")), "#,##0.0;-#,##0.0"),
        (160, "Present value of terminal value", _safe_formula((dcf_years, terminal_fcff, dcf_wacc, dcf_terminal, terminal_discount), f"{terminal_fcff}*(1+{dcf_terminal})/({dcf_wacc}-{dcf_terminal})*{terminal_discount}", invalid_conditions=(*valid_horizon_conditions, f"COUNT(B163:F163)<>{dcf_years}", f"{dcf_wacc}<={dcf_terminal}")), "#,##0.0;-#,##0.0"),
        (161, "Enterprise value", f'={selected_matrix("BG", "dcf")}', "#,##0.0;-#,##0.0"),
        (162, "Net cash / debt", f'=IF(ISNUMBER({selected_net_debt}),{selected_net_debt},"")', "#,##0.0;-#,##0.0"),
        (163, "Equity value", f'={selected_matrix("BH", "dcf")}', "#,##0.0;-#,##0.0"),
        (164, "Diluted shares", f'=IF(ISNUMBER({selected_shares}),{selected_shares},"")', "#,##0.000;-#,##0.000"),
        (165, "Value/share", f'={selected_matrix("BI", "dcf")}', "$0.00;-$0.00"),
    )
    for row, label, formula, fmt in dcf_summary:
        ws[f"G{row}"] = label
        set_formula(ws[f"H{row}"], formula, fmt)
        state_label = "Input ready" if row <= 145 else "Available"
        missing_label = "Manual input required" if row <= 145 else "Unavailable"
        set_formula(ws[f"I{row}"], f'=IF(ISNUMBER(H{row}),"{state_label}","{missing_label}")', "General")
    calculation_rows = (
        (171, "Revenue", 85, "#,##0.0;-#,##0.0"),
        (172, "Operating margin", 88, "0.0%;-0.0%"),
        (173, "Operating income", 89, "#,##0.0;-#,##0.0"),
        (174, "GAAP net income", 92, "#,##0.0;-#,##0.0"),
        (175, "Diluted shares", 95, "#,##0.000;-#,##0.000"),
        (176, "GAAP diluted EPS", 98, "$0.00;-$0.00"),
        (177, "Free cash flow", 94, "#,##0.0;-#,##0.0"),
        (178, "FCF/share", 99, "$0.00;-$0.00"),
        (179, "Net cash / debt", 96, "#,##0.0;-#,##0.0"),
    )
    for row, label, output_row, fmt in calculation_rows:
        ws[f"A{row}"] = label
        set_formula(ws[f"B{row}"], f'=IF(ISNUMBER(B{output_row}),B{output_row},"")', fmt)
        set_formula(ws[f"C{row}"], f'=IF(ISNUMBER({selected_output(output_row)}),{selected_output(output_row)},"")', fmt)

    buyback_rows = (
        (182, "Scenario buyback cash", "Manual assumption", selected_driver(65), "#,##0.0;-#,##0.0"),
        (183, "Execution price", "Manual assumption", selected_driver(66), "$0.00;-$0.00"),
        (184, "Gross shares repurchased", "Buyback cash / execution price", contract._ic_safe_share_delta_formula(selected_driver(65), selected_driver(66), negative=False), "#,##0.000;-#,##0.000"),
        (185, "Share issuance / SBC", "Manual assumption", selected_driver(67), "#,##0.000;-#,##0.000"),
        (186, "Resulting diluted shares", base_shares, selected_output(95), "#,##0.000;-#,##0.000"),
        (
            187,
            "Model-derived GAAP EPS effect",
            _safe_expression((selected_output(92), base_shares), f"{selected_output(92)}/{base_shares}", invalid_conditions=(f"{base_shares}<=0",)),
            contract._ic_safe_eps_difference_formula(f"ISNUMBER({selected_output(98)})", selected_output(98), selected_output(92), base_shares),
            "$0.00;-$0.00",
        ),
        (188, "Historical source-native detail", "", "", "General"),
    )
    for row, label, basis, result, fmt in buyback_rows:
        ws[f"A{row}"] = label
        if basis and (basis == base_shares or str(basis).startswith("IF(")):
            set_formula(ws[f"B{row}"], f'=IF(ISNUMBER({basis}),{basis},"")', fmt)
        else:
            ws[f"B{row}"] = basis
        if result:
            set_formula(ws[f"C{row}"], f"={result}", fmt)
        else:
            set_formula(ws[f"C{row}"], '=""', fmt)

    if "investment_case_sensitivity_formulas" in enabled_formula_ids:
        _apply_sensitivities(
            ws,
            contract,
            assumption_rows,
            selected_output,
            selected_matrix,
        )

    for slot_index, row in enumerate(range(219, 226), start=1):
        slot = f"debate|{slot_index:03d}"
        set_formula(ws[f"A{row}"], f"={support_text(slot, 'Y')}", "General")
        set_formula(ws[f"B{row}"], f"={support_text(slot, 'AO')}", "General")
        state = support_text(slot, "AQ")
        set_formula(
            ws[f"G{row}"],
            (
                f'=IF(A{row}="","",'
                f'IF({state}="manual_review_required","Manual review required",'
                f'IF({state}="accepted","Accepted",'
                f'IF({state}="source_backed","Source-backed","Review required"))))'
            ),
            "General",
        )
        set_formula(ws[f"H{row}"], f'=IF(A{row}="","","Monitor next accepted update and related driver")', "General")
        set_formula(ws[f"L{row}"], f"={support_text(slot, 'BA')}", "General")


def _apply_sensitivities(
    ws: Any,
    contract: Any,
    assumption_rows: dict[str, int],
    selected_output: Any,
    selected_matrix: Any,
) -> None:
    set_formula = contract._set_formula

    families = (
        (
            192,
            193,
            "P/E step (x)",
            "EPS step ($/share)",
            f"$D${assumption_rows['target_pe']}",
            selected_matrix("BE", "pe"),
        ),
        (
            198,
            199,
            "Multiple step (x)",
            "EBITDA step ($m)",
            f"$D${assumption_rows['target_ev_adjusted_ebitda']}",
            selected_matrix("BE", "ev_adjusted_ebitda"),
        ),
        (
            204,
            205,
            "Yield step (%)",
            "FCF step ($m)",
            f"$D${assumption_rows['target_fcf_yield']}",
            selected_matrix("BE", "fcf_yield"),
        ),
        (
            210,
            211,
            "WACC step (%)",
            "Terminal-growth step (%)",
            f"$D${assumption_rows['dcf_wacc']}",
            f"$D${assumption_rows['dcf_terminal_growth']}",
        ),
    )
    for first_row, second_row, first_label, second_label, first_center, second_center in families:
        ws[f"G{first_row}"] = first_label
        ws[f"G{second_row}"] = second_label
        set_formula(
            ws[f"H{first_row}"],
            (
                f'=IF(NOT(ISNUMBER({first_center})),"Center unavailable",'
                f'IF(NOT(ISNUMBER(I{first_row})),"Enter positive step",'
                f'IF(I{first_row}<=0,"Step must be positive","Ready")))'
            ),
            "General",
        )
        set_formula(
            ws[f"H{second_row}"],
            (
                f'=IF(NOT(ISNUMBER({second_center})),"Center unavailable",'
                f'IF(NOT(ISNUMBER(I{second_row})),"Enter positive step",'
                f'IF(I{second_row}<=0,"Step must be positive","Ready")))'
            ),
            "General",
        )

    sensitivity_input_formats = {
        192: "0.00x;-0.00x",
        193: "$0.00;-$0.00",
        198: "0.00x;-0.00x",
        199: "#,##0.0;-#,##0.0",
        204: "0.0%;-0.0%",
        205: "#,##0.0;-#,##0.0",
        210: "0.0%;-0.0%",
        211: "0.0%;-0.0%",
    }
    for row, fmt in sensitivity_input_formats.items():
        ws[f"I{row}"].number_format = fmt

    shares = selected_output(95)
    net_debt = selected_output(96)
    adjusted_ebitda = selected_matrix("BE", "ev_adjusted_ebitda")
    free_cash_flow = selected_matrix("BE", "fcf_yield")
    dcf_years = f"$D${assumption_rows['dcf_forecast_years']}"

    for column_index, column in enumerate("BCD", start=-1):
        set_formula(
            ws[f"{column}194"],
            _safe_formula(
                (f"$D${assumption_rows['target_pe']}", "$I$192"),
                f"$D${assumption_rows['target_pe']}+({column_index})*$I$192",
            ),
            "0.00x;-0.00x",
        )
    for row_index, row in enumerate(range(195, 198), start=-1):
        set_formula(
            ws[f"A{row}"],
            _safe_formula((selected_output(98), "$I$193"), f"{selected_output(98)}+({row_index})*$I$193"),
            "$0.00;-$0.00",
        )
        for column in "BCD":
            set_formula(
                ws[f"{column}{row}"],
                _safe_formula((f"$A{row}", f"{column}$194"), f"$A{row}*{column}$194"),
                "$0.00;-$0.00",
            )

    for column_index, column in enumerate("BCD", start=-1):
        set_formula(
            ws[f"{column}200"],
            _safe_formula(
                (f"$D${assumption_rows['target_ev_adjusted_ebitda']}", "$I$198"),
                f"$D${assumption_rows['target_ev_adjusted_ebitda']}+({column_index})*$I$198",
            ),
            "0.00x;-0.00x",
        )
    for row_index, row in enumerate(range(201, 204), start=-1):
        set_formula(
            ws[f"A{row}"],
            _safe_formula((adjusted_ebitda, "$I$199"), f"{adjusted_ebitda}+({row_index})*$I$199"),
            "#,##0.0;-#,##0.0",
        )
        for column in "BCD":
            set_formula(
                ws[f"{column}{row}"],
                _safe_formula(
                    (f"$A{row}", f"{column}$200", net_debt, shares),
                    f"($A{row}*{column}$200-{net_debt})/{shares}",
                    invalid_conditions=(f"{column}$200<=0", f"{shares}<=0"),
                ),
                "$0.00;-$0.00",
            )


    for column_index, column in enumerate("BCD", start=-1):
        set_formula(
            ws[f"{column}206"],
            _safe_formula(
                (f"$D${assumption_rows['target_fcf_yield']}", "$I$204"),
                f"$D${assumption_rows['target_fcf_yield']}+({column_index})*$I$204",
            ),
            "0.0%;-0.0%",
        )
    for row_index, row in enumerate(range(207, 210), start=-1):
        set_formula(
            ws[f"A{row}"],
            _safe_formula((free_cash_flow, "$I$205"), f"{free_cash_flow}+({row_index})*$I$205"),
            "#,##0.0;-#,##0.0",
        )
        for column in "BCD":
            set_formula(
                ws[f"{column}{row}"],
                _safe_formula(
                    (f"$A{row}", f"{column}$206", shares),
                    f"$A{row}/{column}$206/{shares}",
                    invalid_conditions=(f"{column}$206<=0", f"{shares}<=0"),
                ),
                "$0.00;-$0.00",
            )

    for column_index, column in enumerate("BCD", start=-1):
        set_formula(
            ws[f"{column}212"],
            _safe_formula(
                (f"$D${assumption_rows['dcf_terminal_growth']}", "$I$211"),
                f"$D${assumption_rows['dcf_terminal_growth']}+({column_index})*$I$211",
            ),
            "0.0%;-0.0%",
        )
    for row_index, row in enumerate(range(213, 216), start=-1):
        set_formula(
            ws[f"A{row}"],
            _safe_formula(
                (f"$D${assumption_rows['dcf_wacc']}", "$I$210"),
                f"$D${assumption_rows['dcf_wacc']}+({row_index})*$I$210",
            ),
            "0.0%;-0.0%",
        )
        for column in "BCD":
            wacc = f"$A{row}"
            terminal = f"{column}$212"
            pv_terms = "+".join(
                f"IF({dcf_years}>={year_index},{year_column}$163/(1+{wacc})^{year_index},0)"
                for year_index, year_column in enumerate("BCDEF", start=1)
            )
            terminal_fcff = f"INDEX($B$163:$F$163,1,{dcf_years})"
            expression = (
                f"(({pv_terms})+{terminal_fcff}*(1+{terminal})/"
                f"({wacc}-{terminal})/(1+{wacc})^{dcf_years}-{net_debt})/{shares}"
            )
            set_formula(
                ws[f"{column}{row}"],
                _safe_formula(
                    (dcf_years, wacc, terminal, terminal_fcff, net_debt, shares),
                    expression,
                    invalid_conditions=(
                        f"{dcf_years}<1",
                        f"{dcf_years}>5",
                        f"MOD({dcf_years},1)<>0",
                        f"COUNT($B$163:$F$163)<>{dcf_years}",
                        f"{wacc}<={terminal}",
                        f"{wacc}<=-1",
                        f"{shares}<=0",
                    ),
                ),
                "$0.00;-$0.00",
            )


def configure_investment_case_product_layout(workbook: Any) -> None:
    """Apply the final deterministic Investment Case product layout."""

    from copy import copy

    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import range_boundaries

    from pbi_xbrl.standard_template_formula_contract import user_input_surface_targets

    sheet_name = "{ticker}_Investment_Case"
    if sheet_name not in workbook.sheetnames:
        return
    ws = workbook[sheet_name]

    def overlaps(left: Any, right: str) -> bool:
        left_bounds = (
            int(left.min_col),
            int(left.min_row),
            int(left.max_col),
            int(left.max_row),
        )
        right_bounds = range_boundaries(right)
        return not (
            left_bounds[2] < right_bounds[0]
            or right_bounds[2] < left_bounds[0]
            or left_bounds[3] < right_bounds[1]
            or right_bounds[3] < left_bounds[1]
        )

    def bounds_overlap(
        left: tuple[int, int, int, int],
        right: tuple[int, int, int, int],
    ) -> bool:
        return not (
            left[2] < right[0]
            or right[2] < left[0]
            or left[3] < right[1]
            or right[3] < left[1]
        )

    for merged in tuple(ws.merged_cells.ranges):
        if overlaps(merged, "A13:Q240"):
            ws.unmerge_cells(str(merged))

    section_rows = (13, 38, 71, 82, 103, 119, 129, 141, 152, 167, 190, 217)
    header_rows = (15, 41, 55, 62, 72, 83, 105, 116, 120, 130, 142, 155, 170, 181, 218)
    subsection_rows = (16, 20, 30, 40, 46, 49, 54, 61, 84, 93, 97, 115, 154, 169, 180)
    sensitivity_title_rows = (192, 198, 204, 210)
    narrative_rows = (14, 39, 53, 81, 104, 153, 168, 191)

    # Exact reusable tokens sampled from the useful legacy Investment Case surface.
    title_fill = PatternFill("solid", fgColor="4472C4")
    section_fill = PatternFill("solid", fgColor="5B9BD5")
    subsection_fill = PatternFill("solid", fgColor="DDEBF7")
    header_fill = PatternFill("solid", fgColor="EAF3F8")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    white = PatternFill("solid", fgColor="FFFFFF")
    inactive_fill = PatternFill("solid", fgColor="F2F4F5")
    separator_fill = PatternFill("solid", fgColor="D9E2F3")
    section_font = Font(name="Aptos Display", bold=True, color="FFFFFF", size=14)
    header_font = Font(name="Aptos", bold=True, color="1F2933", size=11)
    subsection_font = Font(name="Aptos", bold=True, color="1F2933", size=11)
    body_font = Font(name="Aptos", color="1F2933", size=11)
    note_font = Font(name="Aptos", color="52606D", italic=True, size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9E2EA"),
        right=Side(style="thin", color="D9E2EA"),
        top=Side(style="thin", color="D9E2EA"),
        bottom=Side(style="thin", color="D9E2EA"),
    )
    body_alignment = Alignment(horizontal="left", vertical="center")
    wrapped_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=13, max_row=225, min_col=1, max_col=13):
        ws.row_dimensions[row[0].row].height = 21.0
        for cell in row:
            cell.fill = copy(white)
            cell.font = copy(body_font)
            cell.border = copy(thin_border)
            cell.alignment = copy(body_alignment)

    for row in section_rows:
        ws.merge_cells(f"A{row}:M{row}")
        cell = ws[f"A{row}"]
        cell.fill = copy(section_fill)
        cell.font = copy(section_font)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = copy(thin_border)
        ws.row_dimensions[row].height = 26.0

    for row in header_rows:
        for column in range(1, 14):
            cell = ws.cell(row, column)
            cell.fill = copy(header_fill)
            cell.font = copy(header_font)
            cell.border = copy(thin_border)
            cell.alignment = copy(header_alignment)
        ws.row_dimensions[row].height = 28.0

    for row in subsection_rows:
        if row in {16, 20, 30, 40, 46, 49, 54, 61, 84, 93, 97, 169, 180}:
            ws.merge_cells(f"A{row}:M{row}")
        for column in range(1, 14):
            cell = ws.cell(row, column)
            cell.fill = copy(subsection_fill)
            cell.font = copy(subsection_font)
            cell.border = copy(thin_border)
            cell.alignment = copy(wrapped_alignment)
        ws.row_dimensions[row].height = 22.0

    narrative_heights = {
        14: 36.0,
        39: 36.0,
        53: 30.0,
        81: 36.0,
        104: 36.0,
        153: 42.0,
        168: 36.0,
        191: 36.0,
    }
    for row in narrative_rows:
        ws.merge_cells(f"A{row}:M{row}")
        anchor = ws[f"A{row}"]
        anchor.font = copy(note_font)
        anchor.alignment = copy(wrapped_alignment)
        anchor.fill = copy(white)
        ws.row_dimensions[row].height = narrative_heights[row]

    # Technical lineage starts in N. The visible product uses controlled prose spans.
    ws.merge_cells("G15:L15")
    for row in (*range(17, 20), *range(21, 30), *range(31, 36)):
        ws.merge_cells(f"G{row}:L{row}")

    for header_row in (41, 55, 62):
        ws.merge_cells(f"F{header_row}:L{header_row}")
    for row in (44, 45, 47, 48, 50, 51, 52, *range(56, 61), *range(63, 69)):
        ws.merge_cells(f"F{row}:L{row}")

    ws.merge_cells("F105:L105")
    for row in range(106, 115):
        ws.merge_cells(f"F{row}:L{row}")
    ws.merge_cells("A115:F115")
    ws.merge_cells("G115:J115")
    ws["G115"].alignment = copy(wrapped_alignment)
    ws.row_dimensions[115].height = 36.0

    ws.merge_cells("C130:E130")
    for row in range(131, 140):
        ws.merge_cells(f"C{row}:E{row}")

    ws.merge_cells("A154:F154")
    ws.merge_cells("G154:M154")

    for first_row in sensitivity_title_rows:
        ws.merge_cells(f"A{first_row}:F{first_row}")
        ws.merge_cells(f"A{first_row + 1}:F{first_row + 1}")
        for column in range(1, 10):
            cell = ws.cell(first_row, column)
            cell.fill = copy(subsection_fill)
            cell.font = copy(subsection_font)
            cell.border = copy(thin_border)
            cell.alignment = copy(wrapped_alignment)
        ws.row_dimensions[first_row].height = 22.0
        ws.row_dimensions[first_row + 1].height = 22.0

    ws.merge_cells("B218:F218")
    ws.merge_cells("H218:K218")
    for row in range(219, 226):
        ws.merge_cells(f"B{row}:F{row}")
        ws.merge_cells(f"H{row}:K{row}")
        ws[f"B{row}"].alignment = copy(wrapped_alignment)
        for column in (1, 7, 8, 12):
            ws.cell(row, column).alignment = copy(wrapped_alignment)

    ws.row_dimensions[43].height = 8.0
    for row in (44, 45, 47, 48, 50, 51, 52, *range(56, 61), *range(63, 69)):
        ws.row_dimensions[row].height = 24.0
    for row in range(73, 81):
        ws.row_dimensions[row].height = 24.0
    for row in (*range(85, 93), *range(94, 97), *range(98, 102)):
        ws.row_dimensions[row].height = 23.0
    for row in range(106, 115):
        ws.row_dimensions[row].height = 24.0
    for row in range(121, 128):
        ws.row_dimensions[row].height = 24.0
    ws.row_dimensions[127].height = 42.0
    for row in range(131, 140):
        ws.row_dimensions[row].height = 24.0
    for row in range(143, 151):
        ws.row_dimensions[row].height = 24.0
    for row in range(156, 166):
        ws.row_dimensions[row].height = 24.0
    for row in (*range(171, 180), *range(182, 189)):
        ws.row_dimensions[row].height = 24.0
    for row in range(219, 226):
        ws.row_dimensions[row].height = 40.0

    for row in range(1, 241):
        ws.row_dimensions[row].hidden = False
        ws.row_dimensions[row].outlineLevel = 0
        ws.row_dimensions[row].collapsed = False
    ws.row_dimensions[12].hidden = True
    for row in range(226, 241):
        ws.row_dimensions[row].hidden = True
    for row in range(221, 226):
        ws.row_dimensions[row].height = 18.0
        ws.row_dimensions[row].hidden = True

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False
    ws.sheet_view.showOutlineSymbols = False
    sensitivity_groups = ((194, 197), (200, 203), (206, 209), (212, 215))
    for start_row, end_row in sensitivity_groups:
        for row in range(start_row, end_row + 1):
            ws.row_dimensions[row].height = 18.0
            ws.row_dimensions[row].hidden = True
            ws.row_dimensions[row].outlineLevel = 0
            ws.row_dimensions[row].collapsed = False

    # B:H deliberately share one width so the dynamic year headers and values align.
    widths = (34.0, 25.0, 25.0, 25.0, 25.0, 25.0, 25.0, 25.0, 18.0, 18.0, 18.0, 18.0, 2.5)
    for column, width in enumerate(widths, start=1):
        dimension = ws.column_dimensions[get_column_letter(column)]
        dimension.width = width
        dimension.hidden = False
        dimension.outlineLevel = 0
        dimension.collapsed = False
    for column in range(14, 54):
        dimension = ws.column_dimensions[get_column_letter(column)]
        dimension.width = 2.0
        dimension.hidden = True
        dimension.outlineLevel = 0
        dimension.collapsed = False

    for range_ref in (
        "A14:M15",
        "A17:M35",
        "A39:M69",
        "A72:M81",
        "A83:M101",
        "A104:M127",
        "A130:M150",
        "A153:M165",
        "A168:M188",
        "A191:M215",
        "A218:M225",
    ):
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.alignment = copy(wrapped_alignment)

    for target in user_input_surface_targets(sheet_name):
        min_col, min_row, max_col, max_row = range_boundaries(target)
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            for cell in row:
                cell.fill = copy(input_fill)
                cell.font = copy(body_font)
                cell.border = copy(thin_border)
                cell.alignment = copy(wrapped_alignment)

    # Headers remain the only centered cells on the visible product.
    for row in header_rows:
        for column in range(1, 14):
            ws.cell(row, column).alignment = copy(header_alignment)

    retained_rules = {}
    for conditional_range, rules in ws.conditional_formatting._cf_rules.items():
        ranges = getattr(conditional_range, "sqref", ())
        if any(
            bounds_overlap(range_boundaries(str(cell_range)), range_boundaries("A38:M68"))
            for cell_range in ranges
        ):
            continue
        retained_rules[conditional_range] = rules
    ws.conditional_formatting._cf_rules.clear()
    ws.conditional_formatting._cf_rules.update(retained_rules)
    for target, group_name in (("A47:L48", "Brand"), ("A50:L52", "Geography")):
        ws.conditional_formatting.add(
            target,
            FormulaRule(
                formula=[f'LOWER($B$42)<>LOWER("{group_name}")'],
                fill=inactive_fill,
            ),
        )
    ws.conditional_formatting.add(
        "A45:L45",
        FormulaRule(
            formula=['LOWER($B$42)<>LOWER("Total Company")'],
            fill=inactive_fill,
        ),
    )

    for coordinate in ("C196", "C202", "C208", "C214"):
        ws[coordinate].fill = copy(separator_fill)
        ws[coordinate].font = copy(header_font)

    for merged in tuple(ws.merged_cells.ranges):
        if overlaps(merged, "A1:M1"):
            ws.unmerge_cells(str(merged))
    for column in range(1, 14):
        ws.cell(1, column).fill = copy(title_fill)
        ws.cell(1, column).border = copy(thin_border)
    ws["A1"].font = Font(name="Aptos Display", bold=True, color="FFFFFF", size=16)
    ws.row_dimensions[1].height = 26.0
    ws["A4"].fill = copy(section_fill)
    ws["A4"].font = Font(name="Aptos Display", bold=True, color="FFFFFF", size=14)
    ws.row_dimensions[4].height = 24.0
    for row in range(5, 12):
        ws.row_dimensions[row].height = 24.0
        for column in range(1, 14):
            ws.cell(row, column).font = copy(body_font)
            ws.cell(row, column).alignment = copy(wrapped_alignment)
    ws.row_dimensions[5].height = 42.0

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    ws.sheet_view.zoomScaleNormal = 110
