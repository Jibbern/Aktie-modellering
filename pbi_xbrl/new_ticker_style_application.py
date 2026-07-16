"""Apply independently reproduced exact-cell style overlays.

The applicator owns no policy decisions and cannot write values.  It changes
only the visual properties present in each planned overlay and verifies that
all other cell properties remain unchanged.
"""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from typing import Any, Mapping

from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Color, PatternFill

from pbi_xbrl.new_ticker_style_planner import StylePlan


class StyleApplicationError(RuntimeError):
    """Raised when an exact style plan cannot be applied without drift."""


@dataclass(frozen=True)
class StyleApplicationResult:
    applied_action_count: int
    styled_cells: tuple[str, ...]


def apply_style_plan(workbook: Any, style_plan: StylePlan) -> StyleApplicationResult:
    """Apply exact planned overlays without changing values or owned structure."""

    if not isinstance(style_plan, StylePlan) or style_plan.status != "PASS":
        raise StyleApplicationError("Style application requires a reproduced PASS StylePlan.")

    seen: set[tuple[str, str]] = set()
    styled_cells: list[str] = []
    for action in style_plan.actions:
        key = (action.sheet, action.cell)
        if key in seen:
            raise StyleApplicationError(f"Style plan changes {action.sheet}!{action.cell} more than once.")
        seen.add(key)
        if action.sheet not in workbook.sheetnames:
            raise StyleApplicationError(f"Style plan references missing sheet {action.sheet!r}.")
        cell = workbook[action.sheet][action.cell]
        if isinstance(cell, MergedCell):
            raise StyleApplicationError(f"Style plan targets non-anchor merged cell {action.sheet}!{action.cell}.")

        before = _protected_cell_snapshot(cell)
        _apply_overlay(cell, action.overlay)
        after = _protected_cell_snapshot(cell)
        if before != after:
            raise StyleApplicationError(
                f"Style overlay changed a non-owned property at {action.sheet}!{action.cell}."
            )
        styled_cells.append(f"{action.sheet}!{action.cell}")

    return StyleApplicationResult(len(styled_cells), tuple(styled_cells))


def _apply_overlay(cell: Any, overlay: Mapping[str, Any]) -> None:
    unknown = set(overlay) - {"fill", "font_color"}
    if unknown:
        raise StyleApplicationError(f"Unsupported style overlay properties: {sorted(unknown)!r}")

    fill = overlay.get("fill")
    if fill is not None:
        if not isinstance(fill, Mapping) or str(fill.get("fill_type") or "") != "solid":
            raise StyleApplicationError("Heatmap overlays require a solid fill contract.")
        color = str(fill.get("fg_color") or "").upper()
        if len(color) != 6 or any(character not in "0123456789ABCDEF" for character in color):
            raise StyleApplicationError(f"Invalid heatmap fill color {color!r}.")
        cell.fill = PatternFill(fill_type="solid", fgColor=color)

    font_color = overlay.get("font_color")
    if font_color is not None:
        color = str(font_color).upper()
        if len(color) != 6 or any(character not in "0123456789ABCDEF" for character in color):
            raise StyleApplicationError(f"Invalid heatmap font color {color!r}.")
        font = copy(cell.font)
        font.color = Color(rgb=color)
        cell.font = font


def _protected_cell_snapshot(cell: Any) -> dict[str, Any]:
    """Capture every property that the fill/font-color overlay does not own."""

    font = copy(cell.font)
    font.color = None
    return {
        "value": cell.value,
        "number_format": cell.number_format,
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "protection": copy(cell.protection),
        "font_without_color": font,
        "comment": (
            (str(cell.comment.author or ""), str(cell.comment.text or ""))
            if cell.comment is not None
            else None
        ),
        "hyperlink": (
            (
                str(cell.hyperlink.target or ""),
                str(cell.hyperlink.location or ""),
                str(cell.hyperlink.display or ""),
                str(cell.hyperlink.tooltip or ""),
            )
            if cell.hyperlink is not None
            else None
        ),
    }
