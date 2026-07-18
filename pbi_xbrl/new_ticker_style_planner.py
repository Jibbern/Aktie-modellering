"""Deterministic exact-cell style planning for the generic ticker template.

The value planner remains the authority for data selection and exact cell
ownership.  This module consumes a completed value plan plus ticker-neutral
style policies and produces fill-only overlays.  It never writes a workbook,
uses cached Excel formula values, or infers behavior from ticker names.
"""
from __future__ import annotations

import hashlib
import json
import math
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from pbi_xbrl.json_schema_validation import load_json_strict, validate_json_schema
from pbi_xbrl.new_ticker_binding_planner import BindingPlan, reproduce_binding_plan
from pbi_xbrl.standard_template_formula_contract import formula_target_contracts
from pbi_xbrl.workbook_modules import (
    binding_owners,
    canonical_json_sha256,
    formula_owners,
    load_workbook_module_manifest,
    resolve_module_profile,
    style_range_contracts,
)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_STYLE_POLICY = ROOT / "docs" / "standard_template_style_policy.json"
STYLE_POLICY_SCHEMA = ROOT / "docs" / "standard_template_style_policy.schema.json"
STYLE_PLAN_SCHEMA = ROOT / "docs" / "new_ticker_style_plan.schema.json"
DEFAULT_MODULE_MANIFEST = ROOT / "docs" / "workbook_module_manifest.json"
DEFAULT_BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
DEFAULT_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
DEFAULT_SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
STYLE_PLAN_VERSION = "1.0.0"
_QUARTER_RE = re.compile(r"^(\d{4})-Q([1-4])$")
_ANNUAL_RE = re.compile(r"^(\d{4})-FY$")
_TRUSTED_STATUSES = {"populated", "source_backed"}


class StylePlanningError(RuntimeError):
    """Raised when the style contract or exact style plan is inconsistent."""


@dataclass(frozen=True)
class EconomicPoint:
    value: float
    unit: str
    source_refs: tuple[str, ...]


@dataclass(frozen=True)
class FormulaEconomicSpec:
    operation: str
    inputs: tuple[str, ...]
    result_unit: str
    period_type: str = "quarter"
    lag: int = 0
    signs: tuple[int, ...] = ()


@dataclass(frozen=True)
class PlannedStyleAction:
    sheet: str
    cell: str
    style_key: str
    policy_id: str
    period: str
    current_value: float
    comparison_period: str | None
    comparison_value: float | None
    signal_value: float
    signal_band: str
    overlay: Mapping[str, Any]
    lineage: tuple[str, ...]

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "cell": self.cell,
            "style_key": self.style_key,
            "policy_id": self.policy_id,
            "period": self.period,
            "current_value": self.current_value,
            "comparison_period": self.comparison_period,
            "comparison_value": self.comparison_value,
            "signal_value": self.signal_value,
            "signal_band": self.signal_band,
            "overlay": dict(self.overlay),
            "lineage": list(self.lineage),
        }


@dataclass(frozen=True)
class StyleDecision:
    sheet: str
    cell: str
    style_key: str
    policy_id: str
    period: str
    applied: bool
    reason: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "cell": self.cell,
            "style_key": self.style_key,
            "policy_id": self.policy_id,
            "period": self.period,
            "applied": self.applied,
            "reason": self.reason,
        }


@dataclass
class StylePlan:
    ticker: str
    module_profile_id: str
    style_contract_digest: str
    binding_plan_digest: str
    actions: list[PlannedStyleAction] = field(default_factory=list)
    decisions: list[StyleDecision] = field(default_factory=list)

    @property
    def status(self) -> str:
        return "PASS"

    def to_dict(self) -> dict[str, Any]:
        return {
            "plan_version": STYLE_PLAN_VERSION,
            "status": self.status,
            "ticker": self.ticker,
            "module_profile_id": self.module_profile_id,
            "style_contract_digest": self.style_contract_digest,
            "binding_plan_digest": self.binding_plan_digest,
            "action_count": len(self.actions),
            "decision_count": len(self.decisions),
            "actions": [action.to_dict() for action in self.actions],
            "decisions": [decision.to_dict() for decision in self.decisions],
        }


def _spec(
    operation: str,
    inputs: Sequence[str],
    result_unit: str,
    *,
    period_type: str = "quarter",
    lag: int = 0,
    signs: Sequence[int] = (),
) -> FormulaEconomicSpec:
    return FormulaEconomicSpec(operation, tuple(inputs), result_unit, period_type, lag, tuple(signs))


# These are economic projections of the existing generic formula contract. They
# supply formula values to the style planner only; they do not write formulas or
# change workbook financial semantics.
FORMULA_ECONOMIC_SPECS: dict[str, FormulaEconomicSpec] = {
    "revenue_ttm": _spec("ttm_sum", ["source:revenue"], "$m"),
    "revenue_yoy": _spec("period_ratio", ["source:revenue"], "%", lag=4),
    "gross_margin": _spec("ratio", ["source:gross_profit", "source:revenue"], "%"),
    "operating_margin": _spec("ratio", ["source:operating_income", "source:revenue"], "%"),
    "operating_margin_ttm": _spec("ratio", ["formula:operating_income_ttm", "formula:revenue_ttm"], "%"),
    "ebitda_margin": _spec("ratio", ["source:base_ebitda", "source:revenue"], "%"),
    "ebitda_yoy": _spec("period_ratio", ["source:base_ebitda"], "%", lag=4),
    "ebitda_ttm": _spec("ttm_sum", ["source:base_ebitda"], "$m"),
    "ebitda_margin_ttm": _spec("ratio", ["formula:ebitda_ttm", "formula:revenue_ttm"], "%"),
    "adjusted_ebitda_ttm": _spec("ttm_sum", ["source:adjusted_ebitda"], "$m"),
    "adjusted_ebitda_delta": _spec("linear", ["source:adjusted_ebitda", "source:base_ebitda"], "$m", signs=[1, -1]),
    "adjusted_ebitda_margin": _spec("ratio", ["source:adjusted_ebitda", "source:revenue"], "%"),
    "adjusted_ebitda_yoy": _spec("period_ratio", ["source:adjusted_ebitda"], "%", lag=4),
    "adjusted_ebitda_margin_ttm": _spec("ratio", ["formula:adjusted_ebitda_ttm", "formula:revenue_ttm"], "%"),
    "operating_income_margin": _spec("ratio", ["source:operating_income", "source:revenue"], "%"),
    "operating_income_ttm": _spec("ttm_sum", ["source:operating_income"], "$m"),
    "operating_income_margin_ttm": _spec("ratio", ["formula:operating_income_ttm", "formula:revenue_ttm"], "%"),
    "net_margin": _spec("ratio", ["source:net_income", "source:revenue"], "%"),
    "net_income_yoy": _spec("period_ratio", ["source:net_income"], "%", lag=4),
    "net_income_ttm": _spec("ttm_sum", ["source:net_income"], "$m"),
    "net_margin_ttm": _spec("ratio", ["formula:net_income_ttm", "formula:revenue_ttm"], "%"),
    "_capex_ttm": _spec("ttm_sum", ["source:capital_expenditures"], "$m"),
    "capex_margin": _spec("ratio", ["source:capital_expenditures", "source:revenue"], "%"),
    "capex_margin_ttm": _spec("ratio", ["formula:_capex_ttm", "formula:revenue_ttm"], "%"),
    "free_cash_flow": _spec("linear", ["source:operating_cash_flow", "source:capital_expenditures"], "$m", signs=[1, -1]),
    "free_cash_flow_yoy_delta": _spec("period_change", ["formula:free_cash_flow"], "$m", lag=4),
    "free_cash_flow_ttm": _spec("ttm_linear", ["source:operating_cash_flow", "source:capital_expenditures"], "$m", signs=[1, -1]),
    "free_cash_flow_ttm_yoy_delta": _spec("period_change", ["formula:free_cash_flow_ttm"], "$m", lag=4),
    "free_cash_flow_margin": _spec("ratio", ["formula:free_cash_flow", "source:revenue"], "%"),
    "free_cash_flow_margin_ttm": _spec("ratio", ["formula:free_cash_flow_ttm", "formula:revenue_ttm"], "%"),
    "buybacks_ttm": _spec("ttm_sum", ["source:buybacks_cash"], "$m"),
    "dividends_ttm": _spec("ttm_sum", ["source:dividends_cash"], "$m"),
    "operating_cash_flow_ttm": _spec("ttm_sum", ["source:operating_cash_flow"], "$m"),
    "net_debt": _spec("linear", ["source:debt_core", "source:cash"], "$m", signs=[1, -1]),
    "net_debt_qoq": _spec("period_change", ["formula:net_debt"], "$m", lag=1),
    "net_debt_yoy": _spec("period_change", ["formula:net_debt"], "$m", lag=4),
    "core_net_cash": _spec("linear", ["source:cash", "source:debt_core"], "$m", signs=[1, -1]),
    "net_cash_with_securities": _spec("linear", ["source:cash", "source:marketable_securities", "source:debt_core"], "$m", signs=[1, 1, -1]),
    "lease_adjusted_net_debt": _spec("linear", ["source:debt_core", "source:lease_liabilities", "source:cash"], "$m", signs=[1, 1, -1]),
    "lease_adjusted_net_debt_with_securities": _spec("linear", ["source:debt_core", "source:lease_liabilities", "source:cash", "source:marketable_securities"], "$m", signs=[1, 1, -1, -1]),
    "net_leverage": _spec("ratio", ["formula:net_debt", "formula:ebitda_ttm"], "x"),
    "adjusted_net_leverage": _spec("ratio", ["formula:net_debt", "formula:adjusted_ebitda_ttm"], "x"),
    "_interest_expense_ttm": _spec("ttm_sum", ["source:interest_expense"], "$m"),
    "_interest_paid_ttm": _spec("ttm_sum", ["source:interest_paid"], "$m"),
    "interest_coverage": _spec("ratio", ["formula:operating_income_ttm", "formula:_interest_expense_ttm"], "x"),
    "cash_interest_coverage": _spec("ratio", ["formula:ebitda_ttm", "formula:_interest_paid_ttm"], "x"),
    "fcf_conversion": _spec("ratio", ["formula:free_cash_flow_ttm", "formula:ebitda_ttm"], "%"),
    "diluted_shares_qoq": _spec("period_change", ["source:diluted_shares"], "m shares", lag=1),
    "diluted_shares_yoy": _spec("period_change", ["source:diluted_shares"], "m shares", lag=4),
    "gaap_eps_yoy": _spec("period_ratio", ["source:eps"], "%", lag=4),
    "gaap_eps_ttm": _spec("ttm_sum", ["source:eps"], "$/share"),
    "adjusted_eps_ttm": _spec("ttm_sum", ["source:adjusted_eps"], "$/share"),
    "book_value_per_share": _spec("ratio", ["source:total_equity", "source:shares_outstanding"], "$/share"),
    "_tangible_equity": _spec("linear", ["source:total_equity", "source:goodwill", "source:intangibles"], "$m", signs=[1, -1, -1]),
    "tangible_book_value_per_share": _spec("ratio", ["formula:_tangible_equity", "source:shares_outstanding"], "$/share"),
    "free_cash_flow_per_share": _spec("ratio", ["formula:free_cash_flow_ttm", "source:diluted_shares"], "$/share"),
    "bs_cash_including_restricted": _spec("sum", ["source:cash", "source:restricted_cash"], "$m"),
    "bs_cash_qoq": _spec("period_change", ["source:cash"], "$m", lag=1),
    "bs_goodwill_assets_ratio": _spec("ratio", ["source:goodwill", "source:total_assets"], "%"),
    "bs_working_capital": _spec("linear", ["source:current_assets", "source:current_liabilities"], "$m", signs=[1, -1]),
    "bs_working_capital_qoq": _spec("period_change", ["formula:bs_working_capital"], "$m", lag=1),
    "bs_current_ratio": _spec("ratio", ["source:current_assets", "source:current_liabilities"], "%"),
    "_quick_assets": _spec("linear", ["source:current_assets", "source:inventory"], "$m", signs=[1, -1]),
    "bs_quick_ratio": _spec("ratio", ["formula:_quick_assets", "source:current_liabilities"], "%"),
    "bs_debt_qoq": _spec("period_change", ["source:debt_core"], "$m", lag=1),
    "bs_inventory_yoy": _spec("period_ratio", ["source:inventory"], "%", lag=4),
    "bs_revenue_yoy": _spec("period_ratio", ["source:revenue"], "%", lag=4),
    "bs_inventory_vs_revenue_growth": _spec("linear", ["formula:bs_inventory_yoy", "formula:bs_revenue_yoy"], "%", signs=[1, -1]),
    "bs_core_net_cash": _spec("linear", ["source:cash", "source:marketable_securities", "source:debt_core"], "$m", signs=[1, 1, -1]),
    "bs_total_lease_liabilities": _spec("sum", ["source:lease_liabilities_current", "source:lease_liabilities_noncurrent"], "$m"),
    "bs_diluted_shares_yoy": _spec("period_ratio", ["source:diluted_shares"], "%", lag=4),
    "annual_gross_margin": _spec("ratio", ["source:gross_profit", "source:revenue"], "%", period_type="fiscal_year"),
    "annual_operating_margin": _spec("ratio", ["source:operating_income", "source:revenue"], "%", period_type="fiscal_year"),
    "annual_ebitda_margin": _spec("ratio", ["source:base_ebitda", "source:revenue"], "%", period_type="fiscal_year"),
    "annual_adjusted_ebitda_margin": _spec("ratio", ["source:adjusted_ebitda", "source:revenue"], "%", period_type="fiscal_year"),
    "annual_net_margin": _spec("ratio", ["source:net_income", "source:revenue"], "%", period_type="fiscal_year"),
    "annual_free_cash_flow": _spec("linear", ["source:operating_cash_flow", "source:capital_expenditures"], "$m", period_type="fiscal_year", signs=[1, -1]),
    "annual_free_cash_flow_margin": _spec("ratio", ["formula:annual_free_cash_flow", "source:revenue"], "%", period_type="fiscal_year"),
    "annual_book_value_per_share": _spec("ratio", ["source:total_equity", "source:shares_outstanding"], "$/share", period_type="fiscal_year"),
    "annual_net_debt": _spec("linear", ["source:debt_core", "source:cash"], "$m", period_type="fiscal_year", signs=[1, -1]),
}


def load_style_policy_contract(
    path: Path | str = DEFAULT_STYLE_POLICY,
    *,
    module_payload: Mapping[str, Any] | None = None,
    binding_payload: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    payload = load_json_strict(Path(path))
    return validate_style_policy_payload(
        payload,
        module_payload=module_payload or load_workbook_module_manifest(DEFAULT_MODULE_MANIFEST),
        binding_payload=binding_payload or load_json_strict(DEFAULT_BINDING_MAP),
    )


def validate_style_policy_payload(
    payload: Mapping[str, Any],
    *,
    module_payload: Mapping[str, Any],
    binding_payload: Mapping[str, Any],
) -> dict[str, Any]:
    if not isinstance(payload, Mapping):
        raise StylePlanningError("Style policy contract must be a JSON object.")
    failures = validate_json_schema(payload, load_json_strict(STYLE_POLICY_SCHEMA))
    if failures:
        sample = "; ".join(f"{field} {keyword}: {message}" for field, keyword, message in failures[:10])
        raise StylePlanningError(f"Style policy contract does not satisfy its schema: {sample}")
    issues = validate_style_policy_contract(
        payload,
        module_payload=module_payload,
        binding_payload=binding_payload,
    )
    if issues:
        raise StylePlanningError("Invalid style policy contract: " + "; ".join(issues[:12]))
    return dict(payload)


def validate_style_policy_contract(
    payload: Mapping[str, Any],
    *,
    module_payload: Mapping[str, Any],
    binding_payload: Mapping[str, Any],
) -> list[str]:
    issues: list[str] = []
    palettes = dict(payload.get("palette_tokens") or {})
    threshold_sets = {str(row.get("threshold_set_id") or ""): row for row in payload.get("threshold_sets") or []}
    policies = [row for row in payload.get("policies") or [] if isinstance(row, Mapping)]
    state_policies = [row for row in payload.get("state_policies") or [] if isinstance(row, Mapping)]
    disabled_targets = [row for row in payload.get("style_disabled") or [] if isinstance(row, Mapping)]
    _add_duplicate_issues(
        issues,
        [str(row.get("policy_id") or "") for row in [*policies, *state_policies]],
        "policy_id",
    )
    _add_duplicate_issues(issues, list(threshold_sets), "threshold_set_id")

    style_ranges = {row.contract_id: row for row in style_range_contracts(module_payload)}
    module_ids = {str(row.get("module_id") or "") for row in module_payload.get("modules") or []}
    module_by_id = {str(row.get("module_id") or ""): row for row in module_payload.get("modules") or []}
    binding_rows = {str(row.get("binding_id") or ""): row for row in binding_payload.get("bindings") or []}
    declared_binding_owners = binding_owners(module_payload)
    declared_formula_owners = formula_owners(module_payload)
    formula_targets = {row.formula_id: row for row in formula_target_contracts()}
    declared_axis_types = _declared_period_axis_types(binding_payload)
    payload_enabled_modules = set(str(value) for value in binding_payload.get("enabled_modules") or module_ids)
    selector_keys: list[str] = []

    for threshold_id, threshold in threshold_sets.items():
        bands = list(threshold.get("bands") or [])
        _add_duplicate_issues(
            issues,
            [str(band.get("band_id") or "") for band in bands],
            f"band_id in threshold set {threshold_id}",
        )
        for band in bands:
            if str(band.get("overlay_token") or "") not in palettes:
                issues.append(f"Threshold set {threshold_id!r} references unknown palette token {band.get('overlay_token')!r}.")
        issues.extend(_validate_threshold_coverage(threshold_id, bands))
        probes = [(-0.15, "strong_negative"), (-0.05, "negative"), (0.0, "neutral"), (0.05, "positive"), (0.15, "strong_positive")]
        for value, expected in probes:
            actual = _band_for_signal(value, threshold)
            if actual != expected:
                issues.append(f"Threshold set {threshold_id!r} maps {value} to {actual!r}, expected {expected!r}.")

    for policy in policies:
        policy_id = str(policy.get("policy_id") or "")
        owner = str(policy.get("owner_module_id") or "")
        policy_is_active = owner in payload_enabled_modules
        if owner not in module_ids:
            issues.append(f"Style policy {policy_id!r} references unknown module {owner!r}.")
        ranges = []
        for style_id in policy.get("owned_style_ids") or []:
            contract = style_ranges.get(str(style_id))
            if contract is None:
                issues.append(f"Style policy {policy_id!r} references unknown style ownership {style_id!r}.")
                continue
            if contract.owner_id != owner:
                issues.append(f"Style policy {policy_id!r} uses style range {style_id!r} owned by {contract.owner_id!r}.")
            ranges.append(contract)
        threshold_id = str(policy.get("threshold_set_id") or "")
        if threshold_id not in threshold_sets:
            issues.append(f"Style policy {policy_id!r} references unknown threshold set {threshold_id!r}.")
        issues.extend(_validate_period_policy(policy))
        for selector in policy.get("target_selectors") or []:
            selector_type = str(selector.get("selector_type") or "")
            target_id = str(selector.get("target_id") or "")
            selector_keys.append(f"{selector_type}:{target_id}")
            applies_to = set(str(value) for value in policy.get("applies_to") or [])
            expected_apply = "raw" if selector_type == "binding" else "formula"
            if expected_apply not in applies_to:
                issues.append(f"Style policy {policy_id!r} selector {target_id!r} is not declared in applies_to.")
            target_ranges: list[tuple[str, str]] = []
            target_owner = ""
            if selector_type == "binding":
                target_owner = declared_binding_owners.get(target_id, "")
                binding = binding_rows.get(target_id)
                if not target_owner:
                    issues.append(f"Style policy {policy_id!r} references unknown binding {target_id!r}.")
                if binding is not None:
                    target_ranges.append((str(binding.get("sheet") or ""), str(binding.get("planner_target") or binding.get("target") or "")))
            elif selector_type == "formula":
                target_owner = declared_formula_owners.get(target_id, "")
                contract = formula_targets.get(target_id)
                if not target_owner or contract is None:
                    issues.append(f"Style policy {policy_id!r} references unknown formula {target_id!r}.")
                elif str(policy.get("comparison_basis") or "") != "disabled" and target_id not in FORMULA_ECONOMIC_SPECS:
                    issues.append(f"Style policy {policy_id!r} formula {target_id!r} lacks deterministic economic evaluation.")
                if contract is not None:
                    target_ranges.extend((contract.sheet, target) for target in contract.targets)
            if policy_is_active and (not target_owner or target_owner in payload_enabled_modules):
                issues.extend(
                    _validate_declared_axis_selector(
                        policy,
                        selector,
                        target_ranges=target_ranges,
                        target_binding=binding_rows.get(target_id) if selector_type == "binding" else None,
                        declared_axis_types=declared_axis_types,
                    )
                )
            if target_owner and target_owner != owner and not _modules_related(owner, target_owner, module_by_id):
                issues.append(
                    f"Style policy {policy_id!r} owner {owner!r} and target owner {target_owner!r} have no declared dependency relationship."
                )
            for sheet, target in target_ranges:
                if target and not any(row.sheet == sheet and _range_contains(row.target, target) for row in ranges):
                    issues.append(f"Style policy {policy_id!r} target {sheet}!{target} is outside its owned style ranges.")

    for policy in state_policies:
        policy_id = str(policy.get("policy_id") or "")
        owner = str(policy.get("owner_module_id") or "")
        policy_is_active = owner in payload_enabled_modules
        if owner not in module_ids:
            issues.append(f"State style policy {policy_id!r} references unknown module {owner!r}.")
        ranges = []
        for style_id in policy.get("owned_style_ids") or []:
            contract = style_ranges.get(str(style_id))
            if contract is None:
                issues.append(f"State style policy {policy_id!r} references unknown style ownership {style_id!r}.")
                continue
            if contract.owner_id != owner:
                issues.append(
                    f"State style policy {policy_id!r} uses style range {style_id!r} owned by {contract.owner_id!r}."
                )
            ranges.append(contract)
        overlays = {str(key): str(value) for key, value in (policy.get("state_overlays") or {}).items()}
        no_style_states = set(map(str, policy.get("no_style_states") or []))
        overlap = sorted(set(overlays) & no_style_states)
        if overlap:
            issues.append(f"State style policy {policy_id!r} both styles and suppresses states {overlap!r}.")
        for state, palette_id in overlays.items():
            if palette_id not in palettes:
                issues.append(
                    f"State style policy {policy_id!r} state {state!r} references unknown palette token {palette_id!r}."
                )
        for selector in policy.get("target_selectors") or []:
            target_id = str(selector.get("target_id") or "")
            selector_keys.append(f"binding_state:{target_id}")
            target_owner = declared_binding_owners.get(target_id, "")
            binding = binding_rows.get(target_id)
            if not target_owner:
                issues.append(f"State style policy {policy_id!r} references unknown binding {target_id!r}.")
                continue
            if binding is None:
                if policy_is_active:
                    issues.append(
                        f"State style policy {policy_id!r} target binding {target_id!r} is absent from the resolved profile."
                    )
                continue
            if target_owner != owner and not _modules_related(owner, target_owner, module_by_id):
                issues.append(
                    f"State style policy {policy_id!r} owner {owner!r} and target owner {target_owner!r} "
                    "have no declared dependency relationship."
                )
            target_ranges, selector_issues = _state_selector_target_ranges(policy, selector, binding)
            issues.extend(selector_issues)
            for sheet, target in target_ranges:
                if not any(row.sheet == sheet and _range_contains(row.target, target) for row in ranges):
                    issues.append(
                        f"State style policy {policy_id!r} target {sheet}!{target} is outside its owned style ranges."
                    )
    _add_duplicate_issues(issues, selector_keys, "style target selector")

    disabled_keys: list[str] = []
    selected_formula_ids = {
        str(selector.get("target_id") or "")
        for policy in policies
        for selector in policy.get("target_selectors") or []
        if str(selector.get("selector_type") or "") == "formula"
    }
    for row in disabled_targets:
        formula_id = str(row.get("formula_id") or "")
        sheet = str(row.get("sheet") or "")
        target = str(row.get("target") or "")
        owner = str(row.get("owner_module_id") or "")
        disabled_keys.append(f"{formula_id}:{sheet}:{target}")
        contract = formula_targets.get(formula_id)
        formula_owner = declared_formula_owners.get(formula_id, "")
        if contract is None or not formula_owner:
            issues.append(f"style_disabled target {formula_id!r} references an unknown formula contract.")
            continue
        if owner != formula_owner:
            issues.append(
                f"style_disabled target {formula_id!r} declares owner {owner!r}, expected formula owner {formula_owner!r}."
            )
        if sheet != contract.sheet or target not in contract.targets:
            issues.append(
                f"style_disabled target {formula_id!r} {sheet}!{target} does not match an exact formula target."
            )
        if not any(
            style_range.owner_id == owner
            and style_range.sheet == sheet
            and _range_contains(style_range.target, target)
            for style_range in style_ranges.values()
        ):
            issues.append(
                f"style_disabled target {formula_id!r} {sheet}!{target} is outside style ranges owned by {owner!r}."
            )
        if formula_id in selected_formula_ids:
            issues.append(
                f"Formula {formula_id!r} is both selected by an active style policy contract and explicitly style_disabled."
            )
    _add_duplicate_issues(issues, disabled_keys, "style_disabled target")
    serialized = _canonical_json(payload).lower()
    for ticker in ("anf", "pbi", "gpre", "gtx"):
        if re.search(rf"\b{ticker}\b", serialized):
            issues.append(f"Style policy contract contains ticker-specific token {ticker.upper()!r}.")
    return issues


def validate_active_style_contract(
    style_contract: Mapping[str, Any],
    *,
    binding_plan: BindingPlan,
    binding_payload: Mapping[str, Any],
    manifest: Mapping[str, Any],
    module_payload: Mapping[str, Any],
) -> list[str]:
    """Validate active axes, target surfaces, and formula-style disposition."""

    issues: list[str] = []
    profile_id = str(binding_payload.get("module_profile_id") or (manifest.get("module_profile") or {}).get("profile_id") or "")
    resolved = resolve_module_profile(module_payload, profile_id)
    enabled = set(resolved.enabled_modules)
    binding_rows = {str(row.get("binding_id") or ""): row for row in binding_payload.get("bindings") or []}
    formula_targets = {row.formula_id: row for row in formula_target_contracts()}
    enabled_formula_ids = set(str(value) for value in (manifest.get("module_profile") or {}).get("enabled_formula_ids") or [])
    declared_binding_owners = binding_owners(module_payload)
    declared_formula_owners = formula_owners(module_payload)
    active_policies = _active_policies_for_profile(style_contract, module_payload, enabled)
    active_state_policies = _active_state_policies_for_profile(style_contract, module_payload, enabled)

    for policy in active_policies:
        axis_id = str(policy.get("period_axis_id") or "")
        axis = binding_plan.period_axes.get(axis_id)
        for selector in policy.get("target_selectors") or []:
            selector_type = str(selector.get("selector_type") or "")
            target_id = str(selector.get("target_id") or "")
            target_owner = (
                declared_binding_owners.get(target_id, "")
                if selector_type == "binding"
                else declared_formula_owners.get(target_id, "")
            )
            if target_owner and target_owner not in enabled:
                continue
            target_ranges: list[tuple[str, str]] = []
            target_period_type = ""
            target_binding = binding_rows.get(target_id) if selector_type == "binding" else None
            if selector_type == "binding":
                if target_binding is not None:
                    target_ranges.append(
                        (
                            str(target_binding.get("sheet") or ""),
                            str(target_binding.get("planner_target") or target_binding.get("target") or ""),
                        )
                    )
                target_owner = declared_binding_owners.get(target_id, "")
                if target_binding is None or not target_owner or target_owner not in enabled:
                    issues.append(
                        _active_target_error(policy, selector, axis_id, target_ranges, "target surface is not active in the resolved profile")
                    )
                    continue
                target_axis_id = str(target_binding.get("period_axis_id") or "")
                if target_axis_id != axis_id:
                    issues.append(
                        _active_target_error(
                            policy,
                            selector,
                            axis_id,
                            target_ranges,
                            f"target binding resolves axis {target_axis_id!r}",
                        )
                    )
                target_period_type = _continuity_period_type(str(target_binding.get("period_axis_continuity") or ""))
            else:
                contract = formula_targets.get(target_id)
                if contract is not None:
                    target_ranges.extend((contract.sheet, target) for target in contract.targets)
                if contract is None or target_id not in enabled_formula_ids:
                    issues.append(
                        _active_target_error(policy, selector, axis_id, target_ranges, "target formula is not active in the resolved profile")
                    )
                    continue
                spec = FORMULA_ECONOMIC_SPECS.get(target_id)
                target_period_type = spec.period_type if spec is not None else ""

            if not isinstance(axis, Mapping):
                issues.append(
                    _active_target_error(policy, selector, axis_id, target_ranges, "axis is absent from the independently reproduced plan")
                )
                continue
            resolved_period_type = _resolved_axis_period_type(axis)
            expected_period_type = str(policy.get("period_type") or "")
            if not resolved_period_type:
                issues.append(
                    _active_target_error(policy, selector, axis_id, target_ranges, "axis has no single valid resolved period type")
                )
            elif resolved_period_type != expected_period_type:
                issues.append(
                    _active_target_error(
                        policy,
                        selector,
                        axis_id,
                        target_ranges,
                        f"resolved axis period type is {resolved_period_type!r}, expected {expected_period_type!r}",
                    )
                )
            if target_period_type and target_period_type != expected_period_type:
                issues.append(
                    _active_target_error(
                        policy,
                        selector,
                        axis_id,
                        target_ranges,
                        f"target period type is {target_period_type!r}, expected {expected_period_type!r}",
                    )
                )

    for policy in active_state_policies:
        for selector in policy.get("target_selectors") or []:
            target_id = str(selector.get("target_id") or "")
            target_owner = declared_binding_owners.get(target_id, "")
            binding = binding_rows.get(target_id)
            if binding is None or not target_owner or target_owner not in enabled:
                issues.append(
                    f"State style policy {policy.get('policy_id')!r} target binding {target_id!r} "
                    "is not active in the resolved profile."
                )
                continue
            _ranges, selector_issues = _state_selector_target_ranges(policy, selector, binding)
            issues.extend(selector_issues)

    issues.extend(
        _validate_active_formula_style_completeness(
            style_contract,
            active_policies=active_policies,
            enabled=enabled,
            enabled_formula_ids=enabled_formula_ids,
            module_payload=module_payload,
        )
    )
    return issues


def style_policy_ids_for_profile(
    style_contract: Mapping[str, Any],
    module_payload: Mapping[str, Any],
    profile_id: str,
) -> tuple[str, ...]:
    """Project style policy IDs from one immutable resolved module profile."""

    resolved = resolve_module_profile(module_payload, profile_id)
    enabled = set(resolved.enabled_modules)
    return tuple(
        sorted(
            {
                str(policy.get("policy_id") or "")
                for policy in [
                    *_active_policies_for_profile(style_contract, module_payload, enabled),
                    *_active_state_policies_for_profile(style_contract, module_payload, enabled),
                ]
            }
        )
    )


def classify_signal_band(value: float, threshold: Mapping[str, Any]) -> str:
    """Classify one normalized signal using the contract's exact boundaries."""

    if not _numeric(value):
        return ""
    return _band_for_signal(float(value), threshold)


def plan_style_actions(
    package: Mapping[str, Any],
    *,
    binding_plan: BindingPlan,
    binding_payload: Mapping[str, Any],
    manifest: Mapping[str, Any],
    module_payload: Mapping[str, Any],
    style_contract: Mapping[str, Any],
) -> StylePlan:
    style_contract = validate_style_policy_payload(
        style_contract,
        module_payload=module_payload,
        binding_payload=binding_payload,
    )
    if not isinstance(binding_plan, BindingPlan) or binding_plan.status != "PASS" or binding_plan.has_blockers:
        raise StylePlanningError("Style planning requires a completed blocker-free BindingPlan.")
    profile_id = str(binding_payload.get("module_profile_id") or (manifest.get("module_profile") or {}).get("profile_id") or "")
    resolved = resolve_module_profile(module_payload, profile_id)
    if tuple(binding_payload.get("enabled_modules") or ()) != resolved.enabled_modules:
        raise StylePlanningError("Binding payload enabled_modules does not match the resolved module profile.")
    expected_profile_signature = canonical_json_sha256(resolved.to_dict())
    if str(binding_payload.get("module_profile_signature") or "") != expected_profile_signature:
        raise StylePlanningError("Binding payload module profile signature differs from the resolved profile.")
    active_contract_issues = validate_active_style_contract(
        style_contract,
        binding_plan=binding_plan,
        binding_payload=binding_payload,
        manifest=manifest,
        module_payload=module_payload,
    )
    if active_contract_issues:
        raise StylePlanningError("Invalid active style contract: " + "; ".join(active_contract_issues[:12]))

    histories, conflicts, segment_histories = _build_histories(package)
    threshold_sets = {str(row["threshold_set_id"]): row for row in style_contract.get("threshold_sets") or []}
    palettes = dict(style_contract.get("palette_tokens") or {})
    enabled = set(resolved.enabled_modules)
    binding_rows = {str(row.get("binding_id") or ""): row for row in binding_payload.get("bindings") or []}
    formula_targets = {row.formula_id: row for row in formula_target_contracts()}
    enabled_formula_ids = set((manifest.get("module_profile") or {}).get("enabled_formula_ids") or [])
    declared_binding_owners = binding_owners(module_payload)
    declared_formula_owners = formula_owners(module_payload)
    writes_by_binding: dict[str, list[Any]] = {}
    for write in binding_plan.planned_writes:
        writes_by_binding.setdefault(write.binding_id, []).append(write)

    evaluator = _FormulaEvaluator(histories, conflicts)
    actions: list[PlannedStyleAction] = []
    decisions: list[StyleDecision] = []
    active_policy_ids = {
        str(policy.get("policy_id") or "")
        for policy in _active_policies_for_profile(style_contract, module_payload, enabled)
    }
    active_state_policy_ids = {
        str(policy.get("policy_id") or "")
        for policy in _active_state_policies_for_profile(style_contract, module_payload, enabled)
    }
    for policy in style_contract.get("policies") or []:
        if str(policy.get("policy_id") or "") not in active_policy_ids:
            continue
        threshold = threshold_sets[str(policy["threshold_set_id"])]
        for selector in policy.get("target_selectors") or []:
            selector_type = str(selector["selector_type"])
            target_id = str(selector["target_id"])
            target_owner = (
                declared_binding_owners.get(target_id, "")
                if selector_type == "binding"
                else declared_formula_owners.get(target_id, "")
            )
            if target_owner and target_owner not in enabled:
                continue
            metric_id = str(selector["metric_id"])
            if selector_type == "binding":
                binding = binding_rows.get(target_id)
                if binding is None:
                    continue
                for write in writes_by_binding.get(target_id, []):
                    if not _numeric(write.value) or str(write.target_type) not in {"number", "pivot_matrix"}:
                        continue
                    period, series_key = _write_period_and_series(write.row_key, metric_id)
                    if not _period_matches(period, str(policy["period_type"])):
                        decisions.append(_decision(write.target_sheet, write.target_cell, target_id, policy, period, False, "period_type_mismatch"))
                        continue
                    current = _binding_point(
                        period,
                        series_key,
                        metric_id,
                        binding,
                        histories,
                        conflicts,
                        segment_histories,
                    )
                    if current is not None and not math.isclose(float(write.value), current.value, rel_tol=1e-9, abs_tol=1e-9):
                        current = None
                        reason = "planned_value_source_mismatch"
                    else:
                        reason = "current_value_missing_or_untrusted"
                    action, decision = _style_decision(
                        sheet=write.target_sheet,
                        cell=write.target_cell,
                        style_key=f"{target_id}|{series_key}|{period}",
                        policy=policy,
                        period=period,
                        current=current,
                        comparison_lookup=lambda comparison_period, sk=series_key, mid=metric_id, b=binding: _binding_point(
                            comparison_period, sk, mid, b, histories, conflicts, segment_histories
                        ),
                        threshold=threshold,
                        palettes=palettes,
                        missing_reason=reason,
                    )
                    decisions.append(decision)
                    if action is not None:
                        actions.append(action)
            else:
                contract = formula_targets.get(target_id)
                axis = binding_plan.period_axes[str(policy["period_axis_id"])]
                if contract is None or target_id not in enabled_formula_ids:
                    continue
                period_to_column = {str(period): str(column) for period, column in (axis.get("period_to_column") or {}).items()}
                for target in contract.targets:
                    min_col, min_row, max_col, max_row = range_boundaries(target)
                    for period, column_letter in period_to_column.items():
                        column = range_boundaries(f"{column_letter}1")[0]
                        if not min_col <= column <= max_col:
                            continue
                        for row in range(min_row, max_row + 1):
                            cell = f"{column_letter}{row}"
                            current, reason = evaluator.evaluate(target_id, period)
                            action, decision = _style_decision(
                                sheet=contract.sheet,
                                cell=cell,
                                style_key=f"{target_id}|{period}",
                                policy=policy,
                                period=period,
                                current=current,
                                comparison_lookup=lambda comparison_period, formula_id=target_id: evaluator.evaluate(formula_id, comparison_period)[0],
                                threshold=threshold,
                                palettes=palettes,
                                missing_reason=reason,
                            )
                            decisions.append(decision)
                            if action is not None:
                                actions.append(action)

    for policy in style_contract.get("state_policies") or []:
        if str(policy.get("policy_id") or "") not in active_state_policy_ids:
            continue
        for selector in policy.get("target_selectors") or []:
            target_id = str(selector.get("target_id") or "")
            target_owner = declared_binding_owners.get(target_id, "")
            if target_owner and target_owner not in enabled:
                continue
            binding = binding_rows.get(target_id)
            if binding is None:
                continue
            state_actions, state_decisions = _plan_binding_state_styles(
                policy,
                selector,
                binding,
                writes_by_binding.get(target_id, []),
                palettes,
            )
            actions.extend(state_actions)
            decisions.extend(state_decisions)

    actions.sort(key=lambda row: (row.sheet, _cell_sort_key(row.cell), row.policy_id, row.style_key))
    decisions.sort(key=lambda row: (row.sheet, _cell_sort_key(row.cell), row.policy_id, row.style_key))
    duplicate_cells = _duplicates([f"{row.sheet}!{row.cell}" for row in actions])
    if duplicate_cells:
        raise StylePlanningError(f"Style plan contains overlapping exact-cell actions: {duplicate_cells[:10]!r}")
    plan = StylePlan(
        ticker=binding_plan.ticker,
        module_profile_id=profile_id,
        style_contract_digest=_payload_digest(style_contract),
        binding_plan_digest=_payload_digest(binding_plan.to_dict()),
        actions=actions,
        decisions=decisions,
    )
    failures = validate_json_schema(plan.to_dict(), load_json_strict(STYLE_PLAN_SCHEMA))
    if failures:
        sample = "; ".join(f"{field} {keyword}: {message}" for field, keyword, message in failures[:10])
        raise StylePlanningError(f"Generated style plan does not satisfy its schema: {sample}")
    return plan


def reproduce_style_plan(
    package: Mapping[str, Any],
    *,
    binding_payload: Mapping[str, Any],
    manifest: Mapping[str, Any],
    shell_path: Path | str = DEFAULT_SHELL,
    module_payload: Mapping[str, Any] | None = None,
    style_contract: Mapping[str, Any] | None = None,
    ticker_override: str | None = None,
    promotion_requested: bool = False,
    expected_binding_plan: Mapping[str, Any] | BindingPlan | None = None,
    expected_style_plan: Mapping[str, Any] | StylePlan | None = None,
) -> tuple[BindingPlan, StylePlan]:
    """Independently reproduce the value plan, then its exact style projection."""

    modules = dict(module_payload or load_workbook_module_manifest(DEFAULT_MODULE_MANIFEST))
    styles = (
        validate_style_policy_payload(
            style_contract,
            module_payload=modules,
            binding_payload=binding_payload,
        )
        if style_contract is not None
        else load_style_policy_contract(module_payload=modules, binding_payload=binding_payload)
    )
    value_plan = reproduce_binding_plan(
        package,
        binding_payload=binding_payload,
        manifest=manifest,
        shell_path=shell_path,
        ticker_override=ticker_override,
        promotion_requested=promotion_requested,
        expected_plan=expected_binding_plan,
    )
    style_plan = plan_style_actions(
        package,
        binding_plan=value_plan,
        binding_payload=binding_payload,
        manifest=manifest,
        module_payload=modules,
        style_contract=styles,
    )
    if expected_style_plan is not None:
        expected = expected_style_plan.to_dict() if isinstance(expected_style_plan, StylePlan) else expected_style_plan
        if not isinstance(expected, Mapping) or _canonical_json(expected) != _canonical_json(style_plan.to_dict()):
            raise StylePlanningError("Cached style plan differs from independent deterministic reproduction.")
    return value_plan, style_plan


def _style_decision(
    *,
    sheet: str,
    cell: str,
    style_key: str,
    policy: Mapping[str, Any],
    period: str,
    current: EconomicPoint | None,
    comparison_lookup: Any,
    threshold: Mapping[str, Any],
    palettes: Mapping[str, Any],
    missing_reason: str,
) -> tuple[PlannedStyleAction | None, StyleDecision]:
    policy_id = str(policy["policy_id"])
    basis = str(policy["comparison_basis"])
    polarity = str(policy["polarity"])
    if basis == "disabled" or polarity == "disabled":
        return None, _decision(sheet, cell, style_key, policy, period, False, "policy_disabled")
    if current is None:
        return None, _decision(sheet, cell, style_key, policy, period, False, missing_reason)
    if current.unit not in set(str(value) for value in policy.get("accepted_units") or []):
        return None, _decision(sheet, cell, style_key, policy, period, False, "current_unit_not_accepted")

    comparison_period: str | None = None
    comparison: EconomicPoint | None = None
    if basis == "direct_value":
        signal = current.value
    else:
        comparison_period = _shift_period(period, int(policy["comparison_lag"]), str(policy["period_type"]))
        if not comparison_period:
            return None, _decision(sheet, cell, style_key, policy, period, False, "comparison_period_invalid")
        comparison = comparison_lookup(comparison_period)
        if comparison is None:
            return None, _decision(sheet, cell, style_key, policy, period, False, "comparator_missing_or_untrusted")
        if current.unit != comparison.unit:
            return None, _decision(sheet, cell, style_key, policy, period, False, "comparator_unit_mismatch")
        if abs(comparison.value) <= 1e-12:
            return None, _decision(sheet, cell, style_key, policy, period, False, "zero_comparator")
        signal = (current.value - comparison.value) / abs(comparison.value)
    if polarity == "lower_is_better":
        signal *= -1.0
    band = _band_for_signal(signal, threshold)
    if not band:
        return None, _decision(sheet, cell, style_key, policy, period, False, "signal_outside_threshold_contract")
    band_row = next(row for row in threshold.get("bands") or [] if str(row.get("band_id") or "") == band)
    overlay = palettes[str(band_row["overlay_token"])]
    lineage = tuple(sorted(set(current.source_refs + (comparison.source_refs if comparison else ()))))
    action = PlannedStyleAction(
        sheet=sheet,
        cell=cell,
        style_key=style_key,
        policy_id=policy_id,
        period=period,
        current_value=current.value,
        comparison_period=comparison_period,
        comparison_value=comparison.value if comparison else None,
        signal_value=signal,
        signal_band=band,
        overlay=dict(overlay),
        lineage=lineage,
    )
    return action, _decision(sheet, cell, style_key, policy, period, True, "applied")


def _decision(
    sheet: str,
    cell: str,
    style_key: str,
    policy: Mapping[str, Any],
    period: str,
    applied: bool,
    reason: str,
) -> StyleDecision:
    return StyleDecision(sheet, cell, style_key, str(policy.get("policy_id") or ""), period, applied, reason)


def _plan_binding_state_styles(
    policy: Mapping[str, Any],
    selector: Mapping[str, Any],
    binding: Mapping[str, Any],
    writes: Sequence[Any],
    palettes: Mapping[str, Any],
) -> tuple[list[PlannedStyleAction], list[StyleDecision]]:
    """Plan categorical state overlays from exact writes in one owned row binding."""

    target_id = str(selector.get("target_id") or "")
    state_field = str(selector.get("state_field") or "")
    period_field = str(selector.get("period_field") or "")
    target_fields = tuple(map(str, selector.get("target_fields") or []))
    column_fields = {
        str(column.get("target_column") or "").upper(): str(column.get("source_field") or column.get("column_id") or "")
        for column in binding.get("target_columns") or []
        if isinstance(column, Mapping)
    }
    grouped: dict[str, dict[str, Any]] = {}
    for write in writes:
        column = re.match(r"^[A-Z]+", str(write.target_cell or ""))
        source_field = column_fields.get(column.group(0) if column else "", "")
        if not source_field:
            continue
        row = grouped.setdefault(str(write.row_key), {})
        if source_field in row:
            raise StylePlanningError(
                f"State style binding {target_id!r} has duplicate planned field {source_field!r} for row {write.row_key!r}."
            )
        row[source_field] = write

    overlays = {str(key): str(value) for key, value in (policy.get("state_overlays") or {}).items()}
    no_style_states = set(map(str, policy.get("no_style_states") or []))
    actions: list[PlannedStyleAction] = []
    decisions: list[StyleDecision] = []
    for row_key in sorted(grouped):
        row = grouped[row_key]
        state_write = row.get(state_field)
        period_write = row.get(period_field)
        if state_write is None or period_write is None:
            raise StylePlanningError(
                f"State style binding {target_id!r} row {row_key!r} lacks {state_field!r} or {period_field!r}."
            )
        state = str(state_write.value or "")
        period = str(period_write.value or "")
        if state not in overlays and state not in no_style_states:
            raise StylePlanningError(
                f"State style policy {policy.get('policy_id')!r} does not classify state {state!r} "
                f"for binding {target_id!r} row {row_key!r}."
            )
        lineage = tuple(sorted({str(write.source_ref) for write in row.values() if str(write.source_ref or "")}))
        for target_field in target_fields:
            target_write = row.get(target_field)
            if target_write is None:
                raise StylePlanningError(
                    f"State style binding {target_id!r} row {row_key!r} lacks styled field {target_field!r}."
                )
            style_key = f"{target_id}|{row_key}|{state}|{target_field}"
            if state in no_style_states:
                decisions.append(
                    _decision(
                        target_write.target_sheet,
                        target_write.target_cell,
                        style_key,
                        policy,
                        period,
                        False,
                        "state_no_style",
                    )
                )
                continue
            overlay = palettes[overlays[state]]
            actions.append(
                PlannedStyleAction(
                    sheet=target_write.target_sheet,
                    cell=target_write.target_cell,
                    style_key=style_key,
                    policy_id=str(policy.get("policy_id") or ""),
                    period=period,
                    current_value=1.0,
                    comparison_period=None,
                    comparison_value=None,
                    signal_value=1.0,
                    signal_band=state,
                    overlay=dict(overlay),
                    lineage=lineage,
                )
            )
            decisions.append(
                _decision(
                    target_write.target_sheet,
                    target_write.target_cell,
                    style_key,
                    policy,
                    period,
                    True,
                    "categorical_state_applied",
                )
            )
    return actions, decisions


class _FormulaEvaluator:
    def __init__(self, histories: Mapping[str, Any], conflicts: set[tuple[str, str, str]]) -> None:
        self.histories = histories
        self.conflicts = conflicts
        self.cache: dict[tuple[str, str], tuple[EconomicPoint | None, str]] = {}

    def evaluate(self, formula_id: str, period: str) -> tuple[EconomicPoint | None, str]:
        key = (formula_id, period)
        if key in self.cache:
            return self.cache[key]
        spec = FORMULA_ECONOMIC_SPECS.get(formula_id)
        if spec is None:
            result = (None, "formula_economics_unavailable")
        elif not _period_matches(period, spec.period_type):
            result = (None, "formula_period_type_mismatch")
        else:
            result = self._evaluate_spec(spec, period)
        self.cache[key] = result
        return result

    def _evaluate_spec(self, spec: FormulaEconomicSpec, period: str) -> tuple[EconomicPoint | None, str]:
        operation = spec.operation
        if operation in {"sum", "linear", "ratio"}:
            points = [self._reference(ref, period, spec.period_type) for ref in spec.inputs]
            if any(point is None for point in points):
                return None, "formula_required_input_missing"
            result = _combine_points(operation, [point for point in points if point is not None], spec)
            return (result, "calculated") if result is not None else (None, "formula_input_unit_or_denominator_invalid")
        if operation in {"ttm_sum", "ttm_linear"}:
            periods = [_shift_period(period, lag, "quarter") for lag in range(4)]
            if any(not item for item in periods) or not _periods_are_consecutive([str(item) for item in periods], "quarter"):
                return None, "formula_ttm_period_coverage_invalid"
            points: list[EconomicPoint] = []
            for item in periods:
                for ref in spec.inputs:
                    point = self._reference(ref, str(item), "quarter")
                    if point is None:
                        return None, "formula_ttm_input_missing"
                    points.append(point)
            result = _combine_ttm_points(points, spec)
            return (result, "calculated") if result is not None else (None, "formula_ttm_unit_or_component_invalid")
        if operation in {"period_change", "period_ratio"}:
            previous_period = _shift_period(period, spec.lag, spec.period_type)
            if not previous_period:
                return None, "formula_comparison_period_invalid"
            current = self._reference(spec.inputs[0], period, spec.period_type)
            previous = self._reference(spec.inputs[0], previous_period, spec.period_type)
            if current is None or previous is None:
                return None, "formula_comparison_input_missing"
            if current.unit != previous.unit:
                return None, "formula_comparison_unit_mismatch"
            if operation == "period_ratio":
                if abs(previous.value) <= 1e-12:
                    return None, "formula_zero_denominator"
                value = (current.value - previous.value) / abs(previous.value)
            else:
                value = current.value - previous.value
            return EconomicPoint(value, spec.result_unit, _merge_refs(current, previous)), "calculated"
        return None, "formula_operation_unsupported"

    def _reference(self, ref: str, period: str, period_type: str) -> EconomicPoint | None:
        kind, _, identifier = ref.partition(":")
        if kind == "source":
            if (period_type, identifier, period) in self.conflicts:
                return None
            return ((self.histories.get(period_type) or {}).get(identifier) or {}).get(period)
        if kind == "formula":
            return self.evaluate(identifier, period)[0]
        return None


class FormulaEconomicLookup:
    """Read-only access to accepted source and formula economic projections.

    Deterministic downstream planners can reuse the formula contract without
    depending on Excel caches or recreating financial calculations.  The facade
    does not expose the evaluator's cache or permit callers to add formulas.
    """

    def __init__(self, package: Mapping[str, Any]) -> None:
        histories, conflicts, _ = _build_histories(package)
        self._histories = histories
        self._evaluator = _FormulaEvaluator(histories, conflicts)

    def formula_point(self, formula_id: str, period: str) -> tuple[EconomicPoint | None, str]:
        return self._evaluator.evaluate(formula_id, period)

    def formula_lineage(self, formula_id: str) -> tuple[str, ...]:
        result: set[str] = set()

        def collect(identifier: str) -> None:
            if identifier in result:
                return
            spec = FORMULA_ECONOMIC_SPECS.get(identifier)
            if spec is None:
                return
            result.add(identifier)
            for ref in spec.inputs:
                kind, _, dependency = ref.partition(":")
                if kind == "formula":
                    collect(dependency)

        collect(formula_id)
        return tuple(sorted(result))

    def periods(self, *, period_type: str = "quarter") -> tuple[str, ...]:
        values = {
            period
            for series in (self._histories.get(period_type) or {}).values()
            for period in series
            if _period_matches(period, period_type)
        }
        return tuple(sorted(values))


def _combine_points(
    operation: str,
    points: Sequence[EconomicPoint],
    spec: FormulaEconomicSpec,
) -> EconomicPoint | None:
    if operation in {"sum", "linear"}:
        if len({point.unit for point in points}) != 1:
            return None
        signs = spec.signs or tuple(1 for _ in points)
        if len(signs) != len(points):
            return None
        value = sum(sign * point.value for sign, point in zip(signs, points))
    else:
        if len(points) != 2 or not _ratio_units_compatible(points[0].unit, points[1].unit, spec.result_unit):
            return None
        if abs(points[1].value) <= 1e-12:
            return None
        value = points[0].value / points[1].value
    return EconomicPoint(value, spec.result_unit, _merge_refs(*points))


def _combine_ttm_points(points: Sequence[EconomicPoint], spec: FormulaEconomicSpec) -> EconomicPoint | None:
    input_count = len(spec.inputs)
    if not input_count or len(points) != input_count * 4:
        return None
    totals: list[EconomicPoint] = []
    for index in range(input_count):
        subset = [points[period_index * input_count + index] for period_index in range(4)]
        if len({point.unit for point in subset}) != 1:
            return None
        totals.append(EconomicPoint(sum(point.value for point in subset), subset[0].unit, _merge_refs(*subset)))
    operation = "linear" if spec.operation == "ttm_linear" else "sum"
    return _combine_points(operation, totals, spec)


def _ratio_units_compatible(numerator: str, denominator: str, result_unit: str) -> bool:
    if result_unit in {"%", "x"}:
        return numerator == denominator
    if result_unit == "$/share":
        return numerator == "$m" and denominator == "m shares"
    return numerator == denominator


def _build_histories(
    package: Mapping[str, Any],
) -> tuple[dict[str, dict[str, dict[str, EconomicPoint]]], set[tuple[str, str, str]], dict[str, dict[str, dict[str, EconomicPoint]]]]:
    histories: dict[str, dict[str, dict[str, EconomicPoint]]] = {"quarter": {}, "fiscal_year": {}}
    conflicts: set[tuple[str, str, str]] = set()
    for row in ((package.get("calculation_history") or {}).get("quarterly_items") or []):
        if not isinstance(row, Mapping):
            continue
        point = _point_from_mapping(row)
        period = str(row.get("period") or "")
        metric = str(row.get("metric") or "")
        if point is not None and _period_matches(period, "quarter") and metric:
            _add_history_point(histories["quarter"], conflicts, "quarter", metric, period, point)
    for row in ((package.get("annual_financials") or {}).get("rows") or []):
        if not isinstance(row, Mapping):
            continue
        period = str(row.get("period") or "")
        if not _period_matches(period, "fiscal_year"):
            continue
        for metric, field in row.items():
            if not isinstance(field, Mapping):
                continue
            point = _point_from_mapping(field)
            if point is not None:
                _add_history_point(histories["fiscal_year"], conflicts, "fiscal_year", str(metric), period, point)

    segment_histories: dict[str, dict[str, dict[str, EconomicPoint]]] = {"quarter": {}, "fiscal_year": {}}
    for row in ((package.get("segments") or {}).get("items") or []):
        if not isinstance(row, Mapping):
            continue
        period = str(row.get("period") or "")
        period_type = "quarter" if _period_matches(period, "quarter") else "fiscal_year" if _period_matches(period, "fiscal_year") else ""
        if not period_type:
            continue
        dimension = str(row.get("dimension") or "")
        member = str(row.get("member") or ((row.get("segment") or {}).get("value") if isinstance(row.get("segment"), Mapping) else "") or "")
        metric = str(row.get("metric") or "revenue")
        series_key = f"{dimension}|{member}|{metric}"
        for field_name in ("revenue", "annual_revenue", "metric_value"):
            field = row.get(field_name)
            point = _point_from_mapping(field) if isinstance(field, Mapping) else None
            if point is not None and (field_name != "metric_value" or metric == "revenue"):
                _add_history_point(segment_histories[period_type], conflicts, f"segment_{period_type}", series_key, period, point)
                break
    return histories, conflicts, segment_histories


def _add_history_point(
    history: dict[str, dict[str, EconomicPoint]],
    conflicts: set[tuple[str, str, str]],
    period_type: str,
    metric: str,
    period: str,
    point: EconomicPoint,
) -> None:
    existing = history.setdefault(metric, {}).get(period)
    if existing is None:
        history[metric][period] = point
        return
    if existing.unit == point.unit and math.isclose(existing.value, point.value, rel_tol=1e-9, abs_tol=1e-9):
        history[metric][period] = EconomicPoint(existing.value, existing.unit, _merge_refs(existing, point))
        return
    history[metric].pop(period, None)
    conflicts.add((period_type, metric, period))


def _point_from_mapping(value: Mapping[str, Any]) -> EconomicPoint | None:
    status = str(value.get("status") or "").strip().lower()
    unit = str(value.get("unit") or "").strip()
    source_ref = str(value.get("source_ref") or "").strip()
    raw = value.get("value")
    if status not in _TRUSTED_STATUSES or not unit or not source_ref or not _numeric(raw):
        return None
    return EconomicPoint(float(raw), unit, (source_ref,))


def _binding_point(
    period: str,
    series_key: str,
    metric_id: str,
    binding: Mapping[str, Any],
    histories: Mapping[str, Any],
    conflicts: set[tuple[str, str, str]],
    segment_histories: Mapping[str, Any],
) -> EconomicPoint | None:
    period_type = "quarter" if _period_matches(period, "quarter") else "fiscal_year" if _period_matches(period, "fiscal_year") else ""
    if not period_type:
        return None
    if str(binding.get("planning_mode") or "") == "pivot_rows":
        conflict_key = (f"segment_{period_type}", series_key, period)
        if conflict_key in conflicts:
            return None
        return ((segment_histories.get(period_type) or {}).get(series_key) or {}).get(period)
    if (period_type, metric_id, period) in conflicts:
        return None
    return ((histories.get(period_type) or {}).get(metric_id) or {}).get(period)


def _write_period_and_series(row_key: str, metric_id: str) -> tuple[str, str]:
    parts = str(row_key or "").split("|")
    period = parts[0] if parts else ""
    return period, "|".join(parts[1:]) if len(parts) > 1 else metric_id


def _band_for_signal(value: float, threshold: Mapping[str, Any]) -> str:
    for band in threshold.get("bands") or []:
        minimum = band.get("minimum")
        maximum = band.get("maximum")
        if minimum is not None:
            if value < float(minimum) or (value == float(minimum) and not bool(band.get("minimum_inclusive"))):
                continue
        if maximum is not None:
            if value > float(maximum) or (value == float(maximum) and not bool(band.get("maximum_inclusive"))):
                continue
        return str(band.get("band_id") or "")
    return ""


def _shift_period(period: str, lag: int, period_type: str) -> str | None:
    if period_type == "quarter":
        match = _QUARTER_RE.fullmatch(period)
        if not match:
            return None
        ordinal = int(match.group(1)) * 4 + int(match.group(2)) - 1 - lag
        return f"{ordinal // 4}-Q{ordinal % 4 + 1}"
    match = _ANNUAL_RE.fullmatch(period)
    return f"{int(match.group(1)) - lag}-FY" if match else None


def _period_matches(period: str, period_type: str) -> bool:
    return bool(_QUARTER_RE.fullmatch(period)) if period_type == "quarter" else bool(_ANNUAL_RE.fullmatch(period))


def _periods_are_consecutive(periods: Sequence[str], period_type: str) -> bool:
    if not periods:
        return False
    for current, previous in zip(periods, periods[1:]):
        if _shift_period(current, 1, period_type) != previous:
            return False
    return True


def _merge_refs(*points: EconomicPoint) -> tuple[str, ...]:
    return tuple(sorted({ref for point in points for ref in point.source_refs if ref}))


def _numeric(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def _declared_period_axis_types(binding_payload: Mapping[str, Any]) -> dict[str, str]:
    axis_types: dict[str, str] = {}
    for binding in binding_payload.get("bindings") or []:
        if str(binding.get("period_axis_role") or "") != "header":
            continue
        axis_id = str(binding.get("period_axis_id") or "")
        period_type = _continuity_period_type(str(binding.get("period_axis_continuity") or ""))
        if axis_id and period_type:
            axis_types[axis_id] = period_type
    return axis_types


def _continuity_period_type(continuity: str) -> str:
    return {
        "consecutive_quarters": "quarter",
        "consecutive_fiscal_years": "fiscal_year",
    }.get(continuity, "")


def _resolved_axis_period_type(axis: Mapping[str, Any]) -> str:
    continuity_type = _continuity_period_type(str(axis.get("continuity") or ""))
    periods = list(axis.get("periods") or (axis.get("period_to_column") or {}).keys())
    observed = {
        "quarter" if _period_matches(str(period), "quarter") else
        "fiscal_year" if _period_matches(str(period), "fiscal_year") else
        ""
        for period in periods
    }
    if "" in observed or len(observed) > 1:
        return ""
    observed_type = next(iter(observed), "")
    if continuity_type and observed_type and continuity_type != observed_type:
        return ""
    return observed_type or continuity_type


def _selector_target_description(
    selector: Mapping[str, Any],
    target_ranges: Sequence[tuple[str, str]],
) -> str:
    selector_type = str(selector.get("selector_type") or "")
    target_id = str(selector.get("target_id") or "")
    locations = ", ".join(f"{sheet}!{target}" for sheet, target in target_ranges if sheet and target)
    target_label = f"{selector_type} {target_id!r}"
    return f"{target_label} ({locations})" if locations else target_label


def _validate_declared_axis_selector(
    policy: Mapping[str, Any],
    selector: Mapping[str, Any],
    *,
    target_ranges: Sequence[tuple[str, str]],
    target_binding: Mapping[str, Any] | None,
    declared_axis_types: Mapping[str, str],
) -> list[str]:
    policy_id = str(policy.get("policy_id") or "")
    axis_id = str(policy.get("period_axis_id") or "")
    expected_period_type = str(policy.get("period_type") or "")
    target_description = _selector_target_description(selector, target_ranges)
    selector_id = f"{selector.get('selector_type')}:{selector.get('target_id')}"
    issues: list[str] = []
    axis_period_type = declared_axis_types.get(axis_id, "")
    if not axis_period_type:
        issues.append(
            f"Style policy {policy_id!r} selector {selector_id!r} has invalid axis {axis_id!r}; "
            f"expected period type {expected_period_type!r}; target {target_description}."
        )
        return issues
    if axis_period_type != expected_period_type:
        issues.append(
            f"Style policy {policy_id!r} selector {selector_id!r} axis {axis_id!r} resolves "
            f"{axis_period_type!r}, expected {expected_period_type!r}; target {target_description}."
        )
    selector_type = str(selector.get("selector_type") or "")
    target_id = str(selector.get("target_id") or "")
    if selector_type == "binding" and target_binding is not None:
        target_axis_id = str(target_binding.get("period_axis_id") or "")
        if target_axis_id != axis_id:
            issues.append(
                f"Style policy {policy_id!r} selector {selector_id!r} axis {axis_id!r} does not match "
                f"target axis {target_axis_id!r}; expected period type {expected_period_type!r}; target {target_description}."
            )
    elif selector_type == "formula":
        spec = FORMULA_ECONOMIC_SPECS.get(target_id)
        if spec is not None and spec.period_type != expected_period_type:
            issues.append(
                f"Style policy {policy_id!r} selector {selector_id!r} axis {axis_id!r} expects "
                f"{expected_period_type!r}, but target formula period type is {spec.period_type!r}; "
                f"target {target_description}."
            )
    return issues


def _active_target_error(
    policy: Mapping[str, Any],
    selector: Mapping[str, Any],
    axis_id: str,
    target_ranges: Sequence[tuple[str, str]],
    reason: str,
) -> str:
    policy_id = str(policy.get("policy_id") or "")
    selector_id = f"{selector.get('selector_type')}:{selector.get('target_id')}"
    expected_period_type = str(policy.get("period_type") or "")
    target_description = _selector_target_description(selector, target_ranges)
    return (
        f"Style policy {policy_id!r} selector {selector_id!r} failed active axis {axis_id!r} "
        f"for expected period type {expected_period_type!r} and target {target_description}: {reason}."
    )


def _validate_active_formula_style_completeness(
    style_contract: Mapping[str, Any],
    *,
    active_policies: Sequence[Mapping[str, Any]],
    enabled: set[str],
    enabled_formula_ids: set[str],
    module_payload: Mapping[str, Any],
) -> list[str]:
    issues: list[str] = []
    formula_targets = {row.formula_id: row for row in formula_target_contracts()}
    formula_owner_by_id = formula_owners(module_payload)
    style_ranges = {row.contract_id: row for row in style_range_contracts(module_payload)}
    active_style_ranges = [row for row in style_ranges.values() if row.owner_id in enabled]
    disabled_rows = [
        row
        for row in style_contract.get("style_disabled") or []
        if str(row.get("owner_module_id") or "") in enabled
        and str(row.get("formula_id") or "") in enabled_formula_ids
    ]

    selected: dict[str, list[tuple[Mapping[str, Any], Mapping[str, Any]]]] = {}
    for policy in active_policies:
        for selector in policy.get("target_selectors") or []:
            if str(selector.get("selector_type") or "") == "formula":
                selected.setdefault(str(selector.get("target_id") or ""), []).append((policy, selector))

    for formula_id in sorted(enabled_formula_ids):
        contract = formula_targets.get(formula_id)
        owner = formula_owner_by_id.get(formula_id, "")
        if contract is None or owner not in enabled:
            continue
        for target in contract.targets:
            covering_style_ranges = [
                row
                for row in active_style_ranges
                if row.sheet == contract.sheet and _ranges_overlap(row.target, target)
            ]
            if not covering_style_ranges:
                continue
            selectors = [
                (policy, selector)
                for policy, selector in selected.get(formula_id, [])
                if any(
                    (style_range := style_ranges.get(str(style_id))) is not None
                    and style_range.sheet == contract.sheet
                    and _range_contains(style_range.target, target)
                    for style_id in policy.get("owned_style_ids") or []
                )
            ]
            dispositions = [
                row
                for row in disabled_rows
                if str(row.get("formula_id") or "") == formula_id
                and str(row.get("sheet") or "") == contract.sheet
                and str(row.get("target") or "") == target
            ]
            total = len(selectors) + len(dispositions)
            if total == 0:
                owned_ranges = ", ".join(
                    f"{row.contract_id}:{row.sheet}!{row.target}" for row in covering_style_ranges
                )
                issues.append(
                    f"Active formula target {formula_id!r} {contract.sheet}!{target} has no style selector or "
                    f"style_disabled disposition inside {owned_ranges}."
                )
            elif total > 1:
                policy_ids = [str(policy.get("policy_id") or "") for policy, _selector in selectors]
                issues.append(
                    f"Active formula target {formula_id!r} {contract.sheet}!{target} has incompatible duplicate "
                    f"style coverage: policies={policy_ids!r}, style_disabled={len(dispositions)}."
                )
    return issues


def _active_policies_for_profile(
    style_contract: Mapping[str, Any],
    module_payload: Mapping[str, Any],
    enabled: set[str],
) -> list[Mapping[str, Any]]:
    binding_owner_by_id = binding_owners(module_payload)
    formula_owner_by_id = formula_owners(module_payload)
    active: list[Mapping[str, Any]] = []
    for policy in style_contract.get("policies") or []:
        if str(policy.get("owner_module_id") or "") not in enabled:
            continue
        selector_owners = {
            (
                binding_owner_by_id.get(str(selector.get("target_id") or ""), "")
                if str(selector.get("selector_type") or "") == "binding"
                else formula_owner_by_id.get(str(selector.get("target_id") or ""), "")
            )
            for selector in policy.get("target_selectors") or []
        }
        if any(owner in enabled for owner in selector_owners if owner):
            active.append(policy)
    return active


def _active_state_policies_for_profile(
    style_contract: Mapping[str, Any],
    module_payload: Mapping[str, Any],
    enabled: set[str],
) -> list[Mapping[str, Any]]:
    binding_owner_by_id = binding_owners(module_payload)
    active: list[Mapping[str, Any]] = []
    for policy in style_contract.get("state_policies") or []:
        if str(policy.get("owner_module_id") or "") not in enabled:
            continue
        selector_owners = {
            binding_owner_by_id.get(str(selector.get("target_id") or ""), "")
            for selector in policy.get("target_selectors") or []
        }
        if any(owner in enabled for owner in selector_owners if owner):
            active.append(policy)
    return active


def _state_selector_target_ranges(
    policy: Mapping[str, Any],
    selector: Mapping[str, Any],
    binding: Mapping[str, Any],
) -> tuple[list[tuple[str, str]], list[str]]:
    policy_id = str(policy.get("policy_id") or "")
    target_id = str(selector.get("target_id") or "")
    target_fields = tuple(map(str, selector.get("target_fields") or []))
    state_field = str(selector.get("state_field") or "")
    period_field = str(selector.get("period_field") or "")
    field_columns = {
        str(column.get("source_field") or column.get("column_id") or ""): str(column.get("target_column") or "").upper()
        for column in binding.get("target_columns") or []
        if isinstance(column, Mapping)
    }
    issues: list[str] = []
    required_fields = [state_field, period_field, *target_fields]
    missing = sorted({field for field in required_fields if not field or field not in field_columns})
    if missing:
        issues.append(
            f"State style policy {policy_id!r} selector {target_id!r} references unmapped binding fields {missing!r}."
        )
    if len(target_fields) != len(set(target_fields)):
        issues.append(f"State style policy {policy_id!r} selector {target_id!r} repeats target_fields.")
    planner_target = str(binding.get("planner_target") or binding.get("target") or "")
    try:
        _left, top, _right, bottom = range_boundaries(planner_target)
    except ValueError:
        return [], [
            *issues,
            f"State style policy {policy_id!r} selector {target_id!r} has invalid binding target {planner_target!r}.",
        ]
    sheet = str(binding.get("sheet") or "")
    ranges = [
        (sheet, f"{field_columns[field]}{top}:{field_columns[field]}{bottom}")
        for field in target_fields
        if field in field_columns
    ]
    return ranges, issues


def _validate_period_policy(policy: Mapping[str, Any]) -> list[str]:
    issues: list[str] = []
    policy_id = str(policy.get("policy_id") or "")
    basis = str(policy.get("comparison_basis") or "")
    lag = int(policy.get("comparison_lag") or 0)
    period_type = str(policy.get("period_type") or "")
    expected = {
        "prior_quarter": ("quarter", 1),
        "prior_year_quarter": ("quarter", 4),
        "prior_ttm": ("quarter", 4),
        "prior_fiscal_year": ("fiscal_year", 1),
        "direct_value": (period_type, 0),
        "disabled": (period_type, 0),
    }.get(basis)
    if expected != (period_type, lag):
        issues.append(f"Style policy {policy_id!r} has incompatible period_type/basis/lag {period_type!r}/{basis!r}/{lag}.")
    if (basis == "disabled") != (str(policy.get("polarity") or "") == "disabled"):
        issues.append(f"Style policy {policy_id!r} must pair disabled basis and polarity.")
    return issues


def _validate_threshold_coverage(threshold_id: str, bands: Sequence[Mapping[str, Any]]) -> list[str]:
    issues: list[str] = []
    if not bands:
        return [f"Threshold set {threshold_id!r} has no bands."]
    if bands[0].get("minimum") is not None:
        issues.append(f"Threshold set {threshold_id!r} does not cover values below its first band.")
    if bands[-1].get("maximum") is not None:
        issues.append(f"Threshold set {threshold_id!r} does not cover values above its last band.")
    for left, right in zip(bands, bands[1:]):
        boundary = left.get("maximum")
        if boundary is None or right.get("minimum") is None or float(boundary) != float(right["minimum"]):
            issues.append(f"Threshold set {threshold_id!r} has a gap or overlap between adjacent band bounds.")
            continue
        inclusive_count = int(bool(left.get("maximum_inclusive"))) + int(bool(right.get("minimum_inclusive")))
        if inclusive_count != 1:
            issues.append(
                f"Threshold set {threshold_id!r} boundary {boundary!r} must belong to exactly one adjacent band."
            )
    return issues


def _modules_related(owner: str, target_owner: str, module_by_id: Mapping[str, Mapping[str, Any]]) -> bool:
    def dependencies(module_id: str) -> set[str]:
        seen: set[str] = set()
        stack = [module_id]
        while stack:
            current = stack.pop()
            for dependency in (module_by_id.get(current) or {}).get("dependencies") or []:
                dependency = str(dependency)
                if dependency not in seen:
                    seen.add(dependency)
                    stack.append(dependency)
        return seen

    return target_owner in dependencies(owner) or owner in dependencies(target_owner)


def _range_contains(outer: str, inner: str) -> bool:
    o_left, o_top, o_right, o_bottom = range_boundaries(outer)
    i_left, i_top, i_right, i_bottom = range_boundaries(inner)
    return o_left <= i_left and i_right <= o_right and o_top <= i_top and i_bottom <= o_bottom


def _ranges_overlap(left: str, right: str) -> bool:
    l_left, l_top, l_right, l_bottom = range_boundaries(left)
    r_left, r_top, r_right, r_bottom = range_boundaries(right)
    return not (l_right < r_left or r_right < l_left or l_bottom < r_top or r_bottom < l_top)


def _add_duplicate_issues(issues: list[str], values: Sequence[str], label: str) -> None:
    for value in _duplicates(values):
        issues.append(f"Duplicate {label} {value!r}.")


def _duplicates(values: Sequence[str]) -> list[str]:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for value in values:
        if value in seen:
            duplicates.add(value)
        seen.add(value)
    return sorted(duplicates)


def _cell_sort_key(cell: str) -> tuple[int, int]:
    column, row = range_boundaries(cell)[:2]
    return row, column


def _canonical_json(value: Any) -> str:
    return json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False)


def _payload_digest(value: Any) -> str:
    return hashlib.sha256(_canonical_json(value).encode("utf-8")).hexdigest()
