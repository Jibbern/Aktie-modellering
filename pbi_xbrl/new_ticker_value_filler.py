"""Value-only executor for an approved generic new-ticker binding plan.

Planning happens in :mod:`pbi_xbrl.new_ticker_binding_planner` before this
module copies or opens an Excel file.  This executor never selects data, walks a
wide range, truncates rows, or combines values for merged cells.  It applies
only exact planned writes after the plan has passed P0/P1 validation.
"""
from __future__ import annotations

import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import range_boundaries

from pbi_xbrl.json_schema_validation import load_json_strict
from pbi_xbrl.new_ticker_binding_planner import (
    BindingPlan,
    BindingPlanReproductionError,
)
from pbi_xbrl.new_ticker_style_application import StyleApplicationError, apply_style_plan
from pbi_xbrl.new_ticker_style_planner import (
    DEFAULT_MODULE_MANIFEST,
    DEFAULT_STYLE_POLICY,
    StylePlanningError,
    reproduce_style_plan,
)
from pbi_xbrl.normalized_company_data_validation import NormalizedDataIssue
from pbi_xbrl.excel_formula_serialization import (
    FormulaSerializationError,
    serialize_workbook_formulas_for_ooxml,
)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = ROOT / "templates" / "standard_stock_model_template.xlsx"
DEFAULT_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
DEFAULT_BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
_QA_SHEETS = {"QA_Log", "Needs_Review", "QA_Checks"}


class NewTickerValueFillerError(RuntimeError):
    """Base error for value-only filler failures."""


class BindingContractError(NewTickerValueFillerError):
    """Raised when a plan or binding violates the frozen shell contract."""


class NormalizedDataValidationError(NewTickerValueFillerError):
    """Raised when planning finds blocking normalized-data issues."""

    def __init__(self, issues: Sequence[NormalizedDataIssue]) -> None:
        self.issues = list(issues)
        rule_ids = ", ".join(issue.rule_id for issue in self.issues[:8])
        super().__init__(f"Blocking normalized-data or binding-plan issues: {rule_ids}")


@dataclass(frozen=True)
class FillResult:
    ticker: str
    output_path: Path
    written_cell_count: int
    styled_cell_count: int
    validation_issue_count: int
    mapping_gap_count: int
    manual_review_count: int


def fill_standard_template_from_package(
    package_path: Path | str,
    *,
    output_path: Path | str,
    ticker_override: str | None = None,
    template_path: Path | str = DEFAULT_TEMPLATE,
    manifest_path: Path | str = DEFAULT_MANIFEST,
    binding_map_path: Path | str = DEFAULT_BINDING_MAP,
    module_manifest_path: Path | str = DEFAULT_MODULE_MANIFEST,
    style_policy_path: Path | str = DEFAULT_STYLE_POLICY,
    promotion_requested: bool = False,
    expected_plan: Mapping[str, Any] | BindingPlan | None = None,
) -> FillResult:
    """Copy the frozen shell and apply an already-safe, exact-cell binding plan.

    All interpretation belongs upstream of this function. Planning and JSON
    Schema validation finish before the template is copied or opened, so a P0/P1
    package problem cannot produce a partial workbook. When ``expected_plan`` is
    supplied, it is comparison-only and must exactly match independent reproduction.
    """

    package = _load_json(Path(package_path))
    manifest = _load_json(Path(manifest_path))
    binding_payload = _load_json(Path(binding_map_path))
    bindings = list(binding_payload.get("bindings") or [])
    ticker = _ticker(package, ticker_override)

    _validate_binding_contract(manifest, bindings)
    try:
        plan, style_plan = reproduce_style_plan(
            package,
            binding_payload=binding_payload,
            manifest=manifest,
            shell_path=Path(template_path),
            module_payload=_load_json(Path(module_manifest_path)),
            style_contract=_load_json(Path(style_policy_path)),
            ticker_override=ticker,
            promotion_requested=promotion_requested,
            expected_binding_plan=expected_plan,
        )
    except BindingPlanReproductionError as exc:
        if exc.plan is not None and exc.plan.has_blockers:
            raise NormalizedDataValidationError(exc.plan.blocking_issues()) from exc
        raise BindingContractError(str(exc)) from exc
    except StylePlanningError as exc:
        raise BindingContractError(str(exc)) from exc

    if ticker != plan.ticker:
        raise BindingContractError(
            f"Ticker override {ticker!r} differs from independently reproduced plan ticker {plan.ticker!r}."
        )
    out_path = Path(output_path)
    if out_path.suffix.lower() != ".xlsx":
        raise NewTickerValueFillerError("Output path must be a macro-free .xlsx file.")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(Path(template_path), out_path)

    wb = load_workbook(out_path, data_only=False, read_only=False)
    try:
        _resolve_ticker_sheet(wb, ticker)
        written = _execute_binding_plan(wb, plan)
        styled = apply_style_plan(wb, style_plan).applied_action_count
        serialize_workbook_formulas_for_ooxml(wb)
        wb.save(out_path)
    except (FormulaSerializationError, StyleApplicationError) as exc:
        raise BindingContractError(str(exc)) from exc
    finally:
        wb.close()

    return FillResult(
        ticker=ticker,
        output_path=out_path,
        written_cell_count=written,
        styled_cell_count=styled,
        validation_issue_count=len(plan.issues),
        mapping_gap_count=len(plan.mapping_gaps),
        manual_review_count=len(plan.manual_review_flags),
    )


def _execute_binding_plan(wb: Any, plan: BindingPlan) -> int:
    """Apply an internally reproduced exact-cell plan.

    This private executor performs no authorization.  The public filler always
    calls :func:`reproduce_binding_plan` immediately before copying the shell.
    """

    if not isinstance(plan, BindingPlan) or plan.status != "PASS" or plan.has_blockers:
        raise BindingContractError("Workbook execution requires an independently reproduced PASS BindingPlan.")
    seen: set[tuple[str, str]] = set()
    written = 0
    for planned_write in plan.planned_writes:
        write = planned_write.to_dict()
        target_sheet = planned_write.target_sheet
        target_cell = planned_write.target_cell
        key = (target_sheet, target_cell)
        if key in seen:
            raise BindingContractError(f"Binding plan writes {target_sheet}!{target_cell} more than once.")
        seen.add(key)
        if target_sheet not in wb.sheetnames:
            raise BindingContractError(f"Binding plan references missing sheet {target_sheet!r}.")
        ws = wb[target_sheet]
        cell = ws[target_cell]
        if isinstance(cell, MergedCell):
            raise BindingContractError(
                f"Binding plan target {target_sheet}!{target_cell} is a non-anchor merged cell. "
                "Each row-schema column must own a distinct writable cell."
            )
        if isinstance(cell.value, str) and cell.value.startswith("="):
            raise BindingContractError(
                f"Binding plan target {target_sheet}!{target_cell} contains a protected formula."
            )
        cell.value = write.get("value")
        written += 1
    return written


def _load_json(path: Path) -> dict[str, Any]:
    payload = load_json_strict(path)
    if not isinstance(payload, dict):
        raise NewTickerValueFillerError(f"JSON contract must be an object: {path}")
    return payload


def _ticker(package: Mapping[str, Any], override: str | None = None) -> str:
    raw = override
    if raw is None:
        meta = package.get("ticker_metadata") if isinstance(package, Mapping) else {}
        ticker_field = meta.get("ticker") if isinstance(meta, Mapping) else ""
        raw = ticker_field.get("value") if isinstance(ticker_field, Mapping) else ticker_field
    ticker = str(raw or "").strip().upper()
    if not ticker:
        raise NewTickerValueFillerError("Ticker is required in the package or --ticker override.")
    if any(ch in ticker for ch in "[]:*?/\\"):
        raise NewTickerValueFillerError(f"Ticker contains characters that are invalid in Excel sheet names: {ticker!r}")
    return ticker


def _resolve_ticker_sheet(wb: Any, ticker: str) -> None:
    token_sheets = [name for name in wb.sheetnames if "{ticker}" in name]
    required_sheet = "{ticker}_Investment_Case"
    resolved_required = required_sheet.replace("{ticker}", ticker)
    if required_sheet not in token_sheets and resolved_required not in wb.sheetnames:
        raise NewTickerValueFillerError("Tokenized investment-case sheet is missing from the frozen shell.")
    for token_sheet in token_sheets:
        resolved = token_sheet.replace("{ticker}", ticker)
        if resolved in wb.sheetnames:
            raise NewTickerValueFillerError(f"Resolved ticker sheet already exists: {resolved!r}.")
        wb[token_sheet].title = resolved
        _replace_defined_name_sheet_token(wb, token_sheet, resolved)
        _replace_formula_sheet_token(wb, token_sheet, resolved)


def _replace_defined_name_sheet_token(wb: Any, token_sheet: str, resolved_sheet: str) -> None:
    quoted_token = f"'{token_sheet}'"
    quoted_resolved = f"'{resolved_sheet}'"
    for name in list(wb.defined_names):
        defined_name = wb.defined_names[name]
        attr_text = getattr(defined_name, "attr_text", None)
        if not isinstance(attr_text, str) or token_sheet not in attr_text:
            continue
        defined_name.attr_text = attr_text.replace(quoted_token, quoted_resolved).replace(token_sheet, resolved_sheet)


def _replace_formula_sheet_token(wb: Any, token_sheet: str, resolved_sheet: str) -> None:
    """Resolve only explicit sheet-name tokens inside formula-owned cells."""

    quoted_token = f"'{token_sheet}'"
    quoted_resolved = f"'{resolved_sheet}'"
    for ws in wb.worksheets:
        # Iterating the rectangular used range materializes blank cells in
        # openpyxl. Restrict token replacement to cells already present in the
        # shell so ticker resolution cannot alter the structural cell surface.
        for cell in tuple(ws._cells.values()):
            value = cell.value
            if not isinstance(value, str) or not value.startswith("=") or token_sheet not in value:
                continue
            cell.value = value.replace(quoted_token, quoted_resolved).replace(token_sheet, resolved_sheet)


def _validate_binding_contract(manifest: Mapping[str, Any], bindings: Sequence[Mapping[str, Any]]) -> None:
    sheets = {str(sheet["sheet"]): sheet for sheet in manifest.get("sheets", []) if isinstance(sheet, Mapping)}
    for binding in bindings:
        if not bool(binding.get("writable")):
            continue
        binding_id = str(binding.get("binding_id") or "")
        sheet_name = str(binding.get("sheet") or "")
        sheet = sheets.get(sheet_name)
        if sheet is None:
            raise BindingContractError(f"Binding references a sheet outside the shell manifest: {sheet_name}")
        target = str(binding.get("target") or "")
        planner_target = str(binding.get("planner_target") or target)
        shell_zone = str(binding.get("shell_zone") or "")
        target_range = _parse_range(target)
        planner_range = _parse_range(planner_target)
        if not _contains(target_range, planner_range):
            raise BindingContractError(f"Binding {binding_id} planner_target {planner_target} is outside declared target {target}.")
        writable_zone = next(
            (zone for zone in sheet.get("writable_zones", []) if isinstance(zone, Mapping) and zone.get("zone_id") == shell_zone),
            None,
        )
        if writable_zone is None:
            raise BindingContractError(f"Binding {binding_id} references missing shell_zone {shell_zone!r}.")
        zone_range = _parse_range(str(writable_zone["target"]))
        if not _contains(zone_range, planner_range):
            raise BindingContractError(f"Binding {binding_id} planner target {planner_target} is outside writable shell zone {shell_zone}.")
        for zone in sheet.get("non_writable_zones", []):
            if isinstance(zone, Mapping) and _overlaps(planner_range, _parse_range(str(zone["target"]))):
                raise BindingContractError(
                    f"Binding {binding_id} planner target {planner_target} overlaps non-writable zone {zone.get('zone_id')}."
                )
        normalized_field = str(binding.get("normalized_field") or "")
        if normalized_field.startswith(("mapping_gaps", "manual_review_flags")) and sheet_name not in _QA_SHEETS:
            raise BindingContractError(f"Binding {binding_id} sends QA data outside its declared QA sheet.")


def _parse_range(range_ref: str) -> tuple[int, int, int, int]:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    except Exception as exc:  # pragma: no cover - openpyxl supplies the exact parse detail
        raise BindingContractError(f"Invalid A1 range {range_ref!r}: {exc}") from exc
    if min_col > max_col or min_row > max_row:
        raise BindingContractError(f"Invalid reversed A1 range {range_ref!r}.")
    return min_col, min_row, max_col, max_row


def _contains(outer: tuple[int, int, int, int], inner: tuple[int, int, int, int]) -> bool:
    outer_left, outer_top, outer_right, outer_bottom = outer
    inner_left, inner_top, inner_right, inner_bottom = inner
    return outer_left <= inner_left and inner_right <= outer_right and outer_top <= inner_top and inner_bottom <= outer_bottom


def _overlaps(first: tuple[int, int, int, int], second: tuple[int, int, int, int]) -> bool:
    f_left, f_top, f_right, f_bottom = first
    s_left, s_top, s_right, s_bottom = second
    return not (f_right < s_left or s_right < f_left or f_bottom < s_top or s_bottom < f_top)
