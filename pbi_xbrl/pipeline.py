"""Thin public pipeline surface used by CLI commands and tests.

The heavy cache-aware assembly work lives in `pipeline_orchestration.py`. This module
keeps the external call shape stable by exposing a small wrapper API for running the
pipeline and packaging the resulting frames into workbook-writer inputs.
"""
from __future__ import annotations

import datetime as dt
import hashlib
import html
import io
import json
import math
import re
import time
from contextlib import contextmanager
from copy import copy
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from datetime import date, datetime, timedelta
try:
    from bs4 import BeautifulSoup
except Exception:  # pragma: no cover - optional dependency in this pipeline
    BeautifulSoup = None

from .debt_parser import build_debt_schedule_tier2, build_debt_tranches_tier2, coerce_number, read_html_tables_any
from .doc_intel import build_doc_intel_outputs, extract_pdf_text_cached, validate_quarter_notes
from .company_profiles import get_company_profile
from .guidance_lexicon import (
    FORWARD_NOTES_LABEL,
    GUIDANCE_UI_METRIC_PRIORITY,
    classify_metric as glx_classify_metric,
    classify_status as glx_classify_status,
    dedup_text_key as glx_dedup_text_key,
    doc_type_priority as glx_doc_type_priority,
    extract_numeric_patterns as glx_extract_numeric_patterns,
    is_preferred_section as glx_is_preferred_section,
    normalize_text as glx_normalize_text,
    normalize_period as glx_normalize_period,
    score_chunk as glx_score_chunk,
    split_sentences as glx_split_sentences,
)
from .metrics import DEBT_TAGS_ORDERED, GAAP_SPECS, MetricSpec, get_income_statement_rules
from .market_data.service import load_market_export_rows
from .non_gaap import build_non_gaap_tier3, find_ex99_docs, infer_quarter_end_from_text, strip_html, parse_adjusted_from_plain_text
from .signals import build_hidden_value_flags, build_hidden_value_outputs, build_signals_base
from .summary_overview import build_company_overview
from .valuation import valuation_engine, valuation_to_frames
from .period_resolver import (
    PickResult,
    _duration_days,
    _filter_unit,
    build_quarter_calendar_from_revenue,
    classify_duration,
    choose_best_tag,
    derive_quarter_from_ytd,
    pick_best_duration,
    pick_best_instant,
    quarter_ends_for_fy,
    self_check_period_logic,
)
from .quarter_notes_lexicon import (
    compact_snippet as qn_compact_snippet,
    is_complete_signal_text as qn_is_complete_signal_text,
    score_promise_candidate as qn_score_promise_candidate,
    score_quarter_note_candidate as qn_score_quarter_note_candidate,
)
from .pdf_utils import silence_pdfminer_warnings
from .sec_xbrl import SecClient, SecConfig, cik10_from_int, cik_from_ticker, companyfacts_to_df, normalize_accession, parse_date
from .validators import info_log_from_audit, needs_review_from_audit, validate_debt_tieout, validate_history
from .pipeline_types import PipelineArtifacts, PipelineConfig, WorkbookInputs


# Revolver scan tuning: keep doc/table selection tight before heavy parsing.
REVOLVER_SCAN_KEYWORDS: Tuple[str, ...] = (
    "revolver",
    "revolving",
    "credit facility",
    "credit agreement",
    "abl",
    "borrowing base",
    "capacity",
    "availability",
    "maturity",
    "covenant",
    "leverage ratio",
    "interest rate",
    "term loan",
    "letters of credit",
    "outstanding borrowings",
)
REVOLVER_DOC_ALLOW_RE = re.compile(
    r"10-?q|10-?k|8-?k|credit|agreement|revolv|facility|debt|liquidity|indenture|amend",
    re.I,
)
REVOLVER_DOC_DENY_RE = re.compile(
    r"compensation|clawback|indemn|stock\s+plan|equity\s+plan|insider|director|governance|"
    r"employment|press\s*release|news\s*release|earnings\s*release|presentation|proxy|consent",
    re.I,
)
REVOLVER_TABLE_HINTS: Tuple[str, ...] = (
    "outstanding",
    "commitment",
    "maturity",
    "interest",
    "rate",
    "facility",
    "availability",
    "borrowing",
    "letters of credit",
)
REVOLVER_TABLE_MAX_CANDIDATES = 15
REVOLVER_DOC_MAX_PER_FILING = 8
REVOLVER_CACHE_VERSION = 4
# Bump whenever stage-level extraction logic changes so stale pickles don't mask fixes.
PIPELINE_STAGE_CACHE_VERSION = 7


def _resolve_path_safe(p: Path) -> Path:
    try:
        return p.resolve()
    except Exception:
        return p


def _path_belongs_to_ticker(
    path_obj: Optional[Path],
    ticker: Optional[str],
    ticker_roots: Optional[List[Path]] = None,
) -> bool:
    if path_obj is None:
        return False
    tkr = str(ticker or "").strip().upper()
    if not tkr:
        return True
    p_res = _resolve_path_safe(Path(path_obj))
    roots = [Path(r) for r in (ticker_roots or []) if r is not None]
    for root in roots:
        r_res = _resolve_path_safe(root)
        try:
            p_res.relative_to(r_res)
            return True
        except Exception:
            continue
    token_re = re.compile(rf"(?<![A-Z0-9]){re.escape(tkr)}(?![A-Z0-9])", re.I)
    p_str = str(p_res)
    if token_re.search(p_str):
        return True
    return False


def _paths_signature(paths: List[Path], max_files: int = 1500) -> str:
    rows: List[str] = []
    for p in sorted([Path(x) for x in paths if x is not None]):
        try:
            st = p.stat()
        except Exception:
            continue
        rows.append(f"{p.name}|{int(st.st_size)}|{int(st.st_mtime)}")
        if len(rows) >= max_files:
            break
    if not rows:
        return "none"
    return hashlib.sha1("||".join(rows).encode("utf-8", errors="ignore")).hexdigest()


def _material_dirs_signature(base_dir: Path, ticker: Optional[str]) -> str:
    t = str(ticker or "").strip().upper()
    dirs = [
        base_dir / "earnings_release",
        base_dir / "Earnings Release",
        base_dir / "Earnings Releases",
        base_dir / "press_release",
        base_dir / "Press Release",
        base_dir / "slides",
        base_dir / "earnings_presentation",
        base_dir / "Earnings Presentation",
        base_dir / "Earnings Transcripts",
        base_dir / "transcripts",
        base_dir / "earnings_transcripts",
        base_dir / "annual_reports",
        base_dir / "financial_statement",
    ]
    if t:
        dirs.extend(
            [
                base_dir / f"{t}-10K",
                base_dir / f"{t}_10K",
                base_dir / f"{t} 10K",
            ]
        )
    files: List[Path] = []
    for d in dirs:
        if not d.exists() or not d.is_dir():
            continue
        try:
            files.extend([p for p in d.rglob("*") if p.is_file()])
        except Exception:
            continue
    return _paths_signature(files)


@contextmanager
def _timed_stage(stage_timings: Dict[str, float], name: str, enabled: bool = True):
    t0 = time.perf_counter()
    try:
        yield
    finally:
        dt_s = time.perf_counter() - t0
        stage_timings[name] = stage_timings.get(name, 0.0) + dt_s
        if enabled:
            print(f"[timing] {name}={dt_s:.2f}s", flush=True)


def _is_quarter_end(d: Optional[dt.date]) -> bool:
    if d is None:
        return False
    return (d.month, d.day) in {(3, 31), (6, 30), (9, 30), (12, 31)}


def _coerce_prev_quarter_end(d: Optional[dt.date]) -> Optional[dt.date]:
    if d is None:
        return None
    # Quarter end for the quarter containing d
    if d.month <= 3:
        q_end = dt.date(d.year, 3, 31)
    elif d.month <= 6:
        q_end = dt.date(d.year, 6, 30)
    elif d.month <= 9:
        q_end = dt.date(d.year, 9, 30)
    else:
        q_end = dt.date(d.year, 12, 31)
    if d >= q_end:
        return q_end
    # Otherwise use the previous quarter end
    if q_end.month == 3:
        return dt.date(d.year - 1, 12, 31)
    if q_end.month == 6:
        return dt.date(d.year, 3, 31)
    if q_end.month == 9:
        return dt.date(d.year, 6, 30)
    return dt.date(d.year, 9, 30)


def _prev_quarter_end_from_qend(d: Optional[dt.date]) -> Optional[dt.date]:
    """Return the previous quarter-end for an already-quarter-end date."""
    if d is None:
        return None
    try:
        ts = pd.Timestamp(d)
        prev = ts - pd.offsets.QuarterEnd(1)
        return prev.date()
    except Exception:
        return None


def _coerce_next_quarter_end(d: Optional[dt.date]) -> Optional[dt.date]:
    if d is None:
        return None
    # Quarter end for the quarter containing d
    if d.month <= 3:
        q_end = dt.date(d.year, 3, 31)
    elif d.month <= 6:
        q_end = dt.date(d.year, 6, 30)
    elif d.month <= 9:
        q_end = dt.date(d.year, 9, 30)
    else:
        q_end = dt.date(d.year, 12, 31)
    if d <= q_end:
        return q_end
    # Otherwise use the next quarter end
    if q_end.month == 3:
        return dt.date(d.year, 6, 30)
    if q_end.month == 6:
        return dt.date(d.year, 9, 30)
    if q_end.month == 9:
        return dt.date(d.year, 12, 31)
    return dt.date(d.year + 1, 3, 31)


def _iter_submission_batches(sec: SecClient, submissions: Dict[str, Any]) -> List[Dict[str, Any]]:
    def _coerce_batch(data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        if not isinstance(data, dict):
            return None
        if "filings" in data and isinstance(data.get("filings"), dict):
            recent = data.get("filings", {}).get("recent", {})
            if recent:
                return recent
        if "accessionNumber" in data and "form" in data:
            return data
        return None

    batches: List[Dict[str, Any]] = []
    base = _coerce_batch(submissions)
    if base:
        batches.append(base)

    files = submissions.get("filings", {}).get("files", []) or []
    for f in files:
        name = f.get("name")
        if not name:
            continue
        url = f"https://data.sec.gov/submissions/{name}"
        try:
            data = sec.get(url, as_json=True, cache_key=f"submissions_{name}")
        except Exception:
            continue
        rec = _coerce_batch(data)
        if rec:
            batches.append(rec)
    return batches


def _build_target_years_from_quarters(target_quarters: Optional[set[dt.date]]) -> Optional[set[int]]:
    if not target_quarters:
        return None
    years = {q.year for q in target_quarters if q is not None}
    if not years:
        return None
    return years | {y + 1 for y in years}


def _dates_match_target_years(
    report_date: Optional[dt.date],
    filing_date: Optional[dt.date],
    target_years: Optional[set[int]],
) -> bool:
    if target_years is None:
        return True
    for d in (report_date, filing_date):
        if d is not None and d.year in target_years:
            return True
    return False


def _make_primary_filing_runtime_cache() -> Dict[str, Any]:
    return {
        "document_bytes": {},
        "html_parse_bundle": {},
        "income_statement_extract": {},
        "ytd_q4_selection": {},
    }


def _load_primary_filing_document_bytes(
    sec: SecClient,
    cik_int: int,
    accn_nd: str,
    doc_name: str,
    filing_runtime_cache: Dict[str, Any],
) -> Optional[bytes]:
    doc_cache = filing_runtime_cache.setdefault("document_bytes", {})
    cache_key = f"{accn_nd}:{str(doc_name).lower()}"
    if cache_key not in doc_cache:
        try:
            doc_cache[cache_key] = sec.download_document(cik_int, accn_nd, doc_name)
        except Exception:
            doc_cache[cache_key] = None
    data = doc_cache.get(cache_key)
    return data if isinstance(data, (bytes, bytearray)) else None


def _parse_primary_filing_html_bundle(html_bytes: bytes) -> Dict[str, Any]:
    html_text = html_bytes.decode("utf-8", errors="ignore")
    return {
        "html": html_text,
        "scale": _detect_scale_from_text(html_text),
        "tables": read_html_tables_any(html_bytes),
    }


def _load_primary_filing_html_bundle(
    sec: SecClient,
    cik_int: int,
    accn_nd: str,
    doc_name: str,
    filing_runtime_cache: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    html_bundle_cache = filing_runtime_cache.setdefault("html_parse_bundle", {})
    cache_key = f"{accn_nd}:{str(doc_name).lower()}"
    if cache_key not in html_bundle_cache:
        html_bytes = _load_primary_filing_document_bytes(sec, cik_int, accn_nd, doc_name, filing_runtime_cache)
        if html_bytes is None:
            html_bundle_cache[cache_key] = None
        else:
            try:
                html_bundle_cache[cache_key] = _parse_primary_filing_html_bundle(html_bytes)
            except Exception:
                html_bundle_cache[cache_key] = None
    bundle = html_bundle_cache.get(cache_key)
    return bundle if isinstance(bundle, dict) else None


def _income_statement_rules_cache_key(rules: Optional[Dict[str, Any]]) -> str:
    if not rules:
        return ""
    try:
        return json.dumps(rules, sort_keys=True, default=str)
    except Exception:
        return repr(sorted((str(k), repr(v)) for k, v in (rules or {}).items()))


def _extract_income_statement_from_primary_doc_cached(
    sec: SecClient,
    cik_int: int,
    accn_nd: str,
    doc_name: str,
    quarter_end: dt.date,
    filing_runtime_cache: Dict[str, Any],
    *,
    rules: Optional[Dict[str, Any]] = None,
    period_hint: str = "3M",
) -> Optional[Dict[str, Any]]:
    extract_cache = filing_runtime_cache.setdefault("income_statement_extract", {})
    rules_key = _income_statement_rules_cache_key(rules)
    cache_key = "|".join(
        [
            accn_nd,
            str(doc_name).lower(),
            str(period_hint or ""),
            quarter_end.isoformat(),
            rules_key,
        ]
    )
    if cache_key not in extract_cache:
        parsed_bundle = _load_primary_filing_html_bundle(sec, cik_int, accn_nd, doc_name, filing_runtime_cache)
        if parsed_bundle is None:
            extract_cache[cache_key] = None
        else:
            try:
                extract_cache[cache_key] = _extract_income_statement_from_html(
                    b"",
                    quarter_end,
                    rules=rules,
                    period_hint=period_hint,
                    parsed_bundle=parsed_bundle,
                )
            except TypeError:
                html_bytes = _load_primary_filing_document_bytes(sec, cik_int, accn_nd, doc_name, filing_runtime_cache)
                if html_bytes is None:
                    extract_cache[cache_key] = None
                else:
                    extract_cache[cache_key] = _extract_income_statement_from_html(
                        html_bytes,
                        quarter_end,
                        rules=rules,
                        period_hint=period_hint,
                    )
    result = extract_cache.get(cache_key)
    return result if isinstance(result, dict) else None


def _prefer_latest_primary_filing_row(candidate: Dict[str, Any], current: Optional[Dict[str, Any]]) -> bool:
    if current is None:
        return True
    cand_filed = pd.to_datetime(candidate.get("filing_date"), errors="coerce")
    cur_filed = pd.to_datetime(current.get("filing_date"), errors="coerce")
    if pd.notna(cand_filed) and pd.notna(cur_filed):
        if cand_filed != cur_filed:
            return bool(cand_filed > cur_filed)
    elif pd.notna(cand_filed) != pd.notna(cur_filed):
        return bool(pd.notna(cand_filed))
    return int(candidate.get("_row_order") or 0) < int(current.get("_row_order") or 0)


def _select_primary_filing_rows_for_ytd_q4(
    filing_rows: List[Dict[str, Any]] | Dict[str, Any],
    target_quarters: Optional[set[dt.date]],
    *,
    filing_runtime_cache: Optional[Dict[str, Any]] = None,
) -> Dict[str, Dict[dt.date, Dict[str, Any]]]:
    if not target_quarters:
        return {
            "fy_rows": {},
            "q3_rows": {},
        }
    fy_targets = {q for q in target_quarters if isinstance(q, dt.date)}
    q3_targets = {
        q3_end
        for fy_end in fy_targets
        for q3_end in [_quarter_ends_from_fy_end(fy_end).get(3)]
        if isinstance(q3_end, dt.date)
    }
    cache_key = tuple(sorted(q.isoformat() for q in fy_targets))
    selection_cache = filing_runtime_cache.setdefault("ytd_q4_selection", {}) if filing_runtime_cache is not None else None
    if selection_cache is not None and cache_key in selection_cache:
        cached = selection_cache[cache_key]
        if isinstance(cached, dict):
            return cached
    fy_rows: Dict[dt.date, Dict[str, Any]] = {}
    q3_rows: Dict[dt.date, Dict[str, Any]] = {}
    if isinstance(filing_rows, dict):
        form_report_index = filing_rows.get("form_report_index") or {}
        q3_index = form_report_index.get("10-Q") or {}
        fy_index = form_report_index.get("10-K") or {}
        q3_rows = {
            q3_end: dict(q3_index[q3_end])
            for q3_end in q3_targets
            if isinstance(q3_end, dt.date) and q3_end.month == 9 and q3_end in q3_index
        }
        fy_rows = {
            fy_end: dict(fy_index[fy_end])
            for fy_end in fy_targets
            if isinstance(fy_end, dt.date) and fy_end in fy_index
        }
    else:
        for row_order, row in enumerate(filing_rows or []):
            form = str(row.get("form") or "").upper().strip()
            q_end = row.get("report_date") or row.get("filing_date")
            if not isinstance(q_end, dt.date):
                continue
            candidate = dict(row)
            candidate["_row_order"] = row_order
            if form in {"10-Q", "10-Q/A"} and q_end in q3_targets and q_end.month == 9:
                if _prefer_latest_primary_filing_row(candidate, q3_rows.get(q_end)):
                    q3_rows[q_end] = candidate
                continue
            if form in {"10-K", "10-K/A"} and q_end in fy_targets:
                if _prefer_latest_primary_filing_row(candidate, fy_rows.get(q_end)):
                    fy_rows[q_end] = candidate
    out = {
        "fy_rows": fy_rows,
        "q3_rows": q3_rows,
    }
    if selection_cache is not None:
        selection_cache[cache_key] = out
    return out


def _build_primary_filing_inventory(
    sec: SecClient,
    submissions: Dict[str, Any],
    *,
    target_years: Optional[set[int]] = None,
) -> Dict[str, Any]:
    rows: List[Dict[str, Any]] = []
    latest_10k_year: Optional[int] = None
    form_report_index: Dict[str, Dict[dt.date, Dict[str, Any]]] = {
        "10-Q": {},
        "10-K": {},
    }
    for batch in _iter_submission_batches(sec, submissions):
        forms = batch.get("form", []) or []
        accns = batch.get("accessionNumber", []) or []
        report_dates = batch.get("reportDate", []) or []
        filing_dates = batch.get("filingDate", []) or []
        primary_docs = batch.get("primaryDocument", []) or []
        n = min(len(forms), len(accns))
        for i in range(n):
            form = str(forms[i] or "").upper().strip()
            accn = accns[i]
            accn_nd = normalize_accession(accn)
            report_raw = report_dates[i] if i < len(report_dates) else None
            filing_raw = filing_dates[i] if i < len(filing_dates) else None
            report_date = parse_date(report_raw)
            filing_date = parse_date(filing_raw)
            if not _dates_match_target_years(report_date, filing_date, target_years):
                continue
            primary_doc = primary_docs[i] if i < len(primary_docs) else None
            row = {
                "form": form,
                "accn": accn,
                "accn_nd": accn_nd,
                "report_date": report_date,
                "report_date_raw": report_raw,
                "filing_date": filing_date,
                "filing_date_raw": filing_raw,
                "primary_doc": primary_doc,
                "_row_order": len(rows),
            }
            rows.append(row)
            form_family = "10-Q" if form in {"10-Q", "10-Q/A"} else ("10-K" if form in {"10-K", "10-K/A"} else "")
            if form_family and isinstance(report_date, dt.date):
                current = form_report_index.setdefault(form_family, {}).get(report_date)
                if _prefer_latest_primary_filing_row(row, current):
                    form_report_index[form_family][report_date] = dict(row)
            if form.startswith("10-K"):
                for d in (report_date, filing_date):
                    if d is not None:
                        latest_10k_year = max(latest_10k_year or d.year, d.year)
                        break
    return {
        "rows": rows,
        "latest_10k_year": latest_10k_year,
        "form_report_index": form_report_index,
    }


def _ensure_primary_filing_inventory(
    sec: SecClient,
    submissions: Dict[str, Any],
    *,
    filing_inventory: Optional[Dict[str, Any]] = None,
    filing_runtime_cache: Optional[Dict[str, Any]] = None,
    target_years: Optional[set[int]] = None,
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    runtime_cache = filing_runtime_cache if filing_runtime_cache is not None else _make_primary_filing_runtime_cache()
    inventory = filing_inventory
    if inventory is None:
        inventory = _build_primary_filing_inventory(sec, submissions, target_years=target_years)
    return inventory, runtime_cache


def _make_ex99_runtime_cache() -> Dict[str, Any]:
    return {
        "accession_index": {},
        "document_bytes": {},
        "legacy_document_bytes": {},
        "index_images": {},
    }


def _load_ex99_accession_index(
    sec: SecClient,
    cik_int: int,
    accn_nd: str,
    ex99_runtime_cache: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    index_cache = ex99_runtime_cache.setdefault("accession_index", {})
    if accn_nd not in index_cache:
        try:
            index_cache[accn_nd] = sec.accession_index_json(cik_int, accn_nd)
        except Exception:
            index_cache[accn_nd] = None
    idx = index_cache.get(accn_nd)
    return idx if isinstance(idx, dict) else None


def _load_ex99_document_bytes(
    sec: SecClient,
    cik_int: int,
    accn_nd: str,
    doc_name: str,
    ex99_runtime_cache: Dict[str, Any],
) -> Optional[bytes]:
    doc_cache = ex99_runtime_cache.setdefault("document_bytes", {})
    cache_key = f"{accn_nd}:{str(doc_name).lower()}"
    if cache_key not in doc_cache:
        try:
            doc_cache[cache_key] = sec.download_document(cik_int, accn_nd, doc_name)
        except Exception:
            doc_cache[cache_key] = None
    data = doc_cache.get(cache_key)
    return data if isinstance(data, (bytes, bytearray)) else None


def _load_legacy_ex99_document_bytes(
    path_in: Path,
    ex99_runtime_cache: Dict[str, Any],
) -> Optional[bytes]:
    legacy_cache = ex99_runtime_cache.setdefault("legacy_document_bytes", {})
    cache_key = str(_resolve_path_safe(path_in))
    if cache_key not in legacy_cache:
        try:
            legacy_cache[cache_key] = path_in.read_bytes()
        except Exception:
            legacy_cache[cache_key] = None
    data = legacy_cache.get(cache_key)
    return data if isinstance(data, (bytes, bytearray)) else None


def _load_ex99_index_images(
    sec: SecClient,
    cik_int: int,
    accn_nd: str,
    idx: Dict[str, Any],
    ex99_runtime_cache: Dict[str, Any],
) -> List[Any]:
    image_cache = ex99_runtime_cache.setdefault("index_images", {})
    if accn_nd not in image_cache:
        try:
            image_cache[accn_nd] = sec.download_index_images(cik_int, accn_nd, idx) or []
        except Exception:
            image_cache[accn_nd] = []
    images = image_cache.get(accn_nd) or []
    return list(images)


def _collect_legacy_ex99_cache_entries(sec: SecClient) -> List[Dict[str, Any]]:
    cache_dir = getattr(sec, "cache_dir", None)
    if cache_dir is None:
        return []
    entries: List[Dict[str, Any]] = []
    for path_in in sorted(Path(cache_dir).glob("doc_*ex99*"), key=lambda p: p.name.lower()):
        if not path_in.is_file():
            continue
        name_lower = path_in.name.lower()
        match = re.match(r"doc_(\d{18})_(.+)", path_in.name)
        entries.append(
            {
                "path": path_in,
                "name": path_in.name,
                "accn_nd": match.group(1) if match else None,
                "doc": match.group(2) if match else path_in.name,
                "is_html_like": ".htm" in name_lower,
                "is_pdf": name_lower.endswith(".pdf"),
            }
        )
    return entries


def _build_ex99_accession_inventory(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    *,
    ex99_runtime_cache: Optional[Dict[str, Any]] = None,
    target_years: Optional[set[int]] = None,
) -> Dict[str, Any]:
    runtime_cache = ex99_runtime_cache if ex99_runtime_cache is not None else _make_ex99_runtime_cache()
    accn_dates: Dict[str, Tuple[Optional[dt.date], Optional[dt.date]]] = {}
    eight_k_rows: List[Dict[str, Any]] = []

    for batch in _iter_submission_batches(sec, submissions):
        forms = batch.get("form", []) or []
        accns = batch.get("accessionNumber", []) or []
        filing_dates = batch.get("filingDate", []) or []
        report_dates = batch.get("reportDate", []) or []
        n = min(len(forms), len(accns))
        for i in range(n):
            accn = accns[i]
            accn_nd = normalize_accession(accn)
            filing_raw = filing_dates[i] if i < len(filing_dates) else None
            report_raw = report_dates[i] if i < len(report_dates) else None
            filing_date = parse_date(filing_raw)
            report_date = parse_date(report_raw)
            accn_dates[accn_nd] = (report_date, filing_date)
            if forms[i] != "8-K":
                continue
            if not _dates_match_target_years(report_date, filing_date, target_years):
                continue
            idx = _load_ex99_accession_index(sec, cik_int, accn_nd, runtime_cache)
            if idx is None:
                continue
            exdocs = list(find_ex99_docs(idx) or [])
            eight_k_rows.append(
                {
                    "accn": accn,
                    "accn_nd": accn_nd,
                    "report_date": report_date,
                    "report_date_raw": report_raw,
                    "filing_date": filing_date,
                    "filing_date_raw": filing_raw,
                    "exdocs": exdocs,
                    "is_image_only": not bool(exdocs),
                }
            )

    return {
        "accn_dates": accn_dates,
        "eight_k_rows": eight_k_rows,
        "legacy_docs": _collect_legacy_ex99_cache_entries(sec),
    }


def _ensure_ex99_inventory(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    *,
    ex99_inventory: Optional[Dict[str, Any]] = None,
    ex99_runtime_cache: Optional[Dict[str, Any]] = None,
    target_years: Optional[set[int]] = None,
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    runtime_cache = ex99_runtime_cache if ex99_runtime_cache is not None else _make_ex99_runtime_cache()
    inventory = ex99_inventory
    if inventory is None:
        inventory = _build_ex99_accession_inventory(
            sec,
            cik_int,
            submissions,
            ex99_runtime_cache=runtime_cache,
            target_years=target_years,
        )
    return inventory, runtime_cache


def _pick_instant_tag(df_all: pd.DataFrame, end: dt.date, tag: str, prefer_forms: List[str]) -> Optional[pd.Series]:
    s = df_all[(df_all["tag"] == tag) & df_all["end_d"].notna()].copy()
    if s.empty:
        return None
    s = s[s["end_d"] == end].copy()
    if s.empty:
        return None
    u = s["unit"].astype(str).str.upper()
    s = s[u.str.startswith("USD", na=False) & ~u.str.contains("/", na=False)].copy()
    if s.empty:
        return None
    return pick_best_instant(s, end=end, prefer_forms=prefer_forms)


def choose_total_debt(
    noncurrent: Optional[float],
    current: Optional[float],
    longtermdebt: Optional[float],
) -> Tuple[Optional[float], str]:
    """
    Pick a debt total without double-counting current maturities.
    """
    if noncurrent is not None and current is not None:
        return float(noncurrent + current), "noncurrent_plus_current"
    if longtermdebt is not None:
        return float(longtermdebt), "longterm_includes_current_or_total"
    if noncurrent is not None:
        return float(noncurrent), "noncurrent_only"
    if current is not None:
        return float(current), "current_only"
    return None, "missing"


def _latest_rec(records: List[pd.Series]) -> Optional[pd.Series]:
    if not records:
        return None
    ranked = sorted(
        records,
        key=lambda r: (
            pd.to_datetime(r.get("filed_d"), errors="coerce")
            if pd.notna(pd.to_datetime(r.get("filed_d"), errors="coerce"))
            else pd.Timestamp("1900-01-01")
        ),
    )
    return ranked[-1] if ranked else None


def compute_total_debt_instant(df_all: pd.DataFrame, end: dt.date, prefer_forms: List[str]) -> Optional[PickResult]:
    """
    Compute total debt:
    - If 'Debt' exists -> pick best instant for that tag.
    - Else use components without double-counting current maturities.
    """
    rec_total = _pick_instant_tag(df_all, end=end, tag="Debt", prefer_forms=prefer_forms)
    if rec_total is not None:
        return PickResult(
            value=float(rec_total["val"]),
            source="direct",
            source_choice="debt_tag",
            tag="Debt",
            accn=str(rec_total["accn"]),
            form=str(rec_total["form"]),
            filed=rec_total["filed_d"],
            start=rec_total["start_d"],
            end=rec_total["end_d"],
            unit=str(rec_total["unit"]),
            duration_days=None,
            note="total_debt from Debt tag",
        )

    rec_non = _pick_first_instant_tag(
        df_all,
        end=end,
        tags=["LongTermDebtNoncurrent", "LongTermDebtAndCapitalLeaseObligations"],
        prefer_forms=prefer_forms,
    )
    rec_long = _pick_instant_tag(df_all, end=end, tag="LongTermDebt", prefer_forms=prefer_forms)
    rec_cur = _pick_first_instant_tag(
        df_all,
        end=end,
        tags=["LongTermDebtCurrent", "DebtCurrent"],
        prefer_forms=prefer_forms,
    )
    non_val = float(rec_non["val"]) if rec_non is not None and pd.notna(rec_non.get("val")) else None
    long_val = float(rec_long["val"]) if rec_long is not None and pd.notna(rec_long.get("val")) else None
    cur_val = float(rec_cur["val"]) if rec_cur is not None and pd.notna(rec_cur.get("val")) else None
    value, source_choice = choose_total_debt(non_val, cur_val, long_val)
    if value is None:
        return None

    used_recs: List[pd.Series] = []
    used_tags: List[str] = []
    note = ""
    source = "direct_component"
    if source_choice == "noncurrent_plus_current":
        if rec_non is not None:
            used_recs.append(rec_non)
            used_tags.append(str(rec_non.get("tag")))
        if rec_cur is not None:
            used_recs.append(rec_cur)
            used_tags.append(str(rec_cur.get("tag")))
        source = "derived_parts"
        note = "total_debt = noncurrent + current debt (preferred)"
    elif source_choice == "longterm_includes_current_or_total":
        if rec_long is not None:
            used_recs.append(rec_long)
            used_tags.append(str(rec_long.get("tag")))
        note = "total_debt from LongTermDebt (already includes current for some issuers)"
    elif source_choice == "noncurrent_only":
        if rec_non is not None:
            used_recs.append(rec_non)
            used_tags.append(str(rec_non.get("tag")))
        note = "total_debt from noncurrent debt only (current portion unavailable)"
    else:
        if rec_cur is not None:
            used_recs.append(rec_cur)
            used_tags.append(str(rec_cur.get("tag")))
        note = "total_debt fallback to current debt only"

    last_rec = _latest_rec(used_recs)

    return PickResult(
        value=float(value),
        source=source,
        source_choice=source_choice,
        tag=",".join(used_tags),
        accn=str(last_rec["accn"]) if last_rec is not None else None,
        form=str(last_rec["form"]) if last_rec is not None else None,
        filed=last_rec["filed_d"] if last_rec is not None else None,
        start=last_rec["start_d"] if last_rec is not None else None,
        end=last_rec["end_d"] if last_rec is not None else None,
        unit=str(last_rec["unit"]) if last_rec is not None else None,
        duration_days=None,
        note=note,
    )


def _pick_first_instant_tag(
    df_all: pd.DataFrame,
    end: dt.date,
    tags: List[str],
    prefer_forms: List[str],
) -> Optional[pd.Series]:
    for tag in tags:
        rec = _pick_instant_tag(df_all, end=end, tag=tag, prefer_forms=prefer_forms)
        if rec is not None:
            return rec
    return None


def compute_debt_core_instant(df_all: pd.DataFrame, end: dt.date, prefer_forms: List[str]) -> Optional[PickResult]:
    rec_total = _pick_instant_tag(df_all, end=end, tag="Debt", prefer_forms=prefer_forms)
    if rec_total is not None:
        return PickResult(
            value=float(rec_total["val"]),
            source="direct",
            source_choice="debt_tag",
            tag="Debt",
            accn=str(rec_total["accn"]),
            form=str(rec_total["form"]),
            filed=rec_total["filed_d"],
            start=rec_total["start_d"],
            end=rec_total["end_d"],
            unit=str(rec_total["unit"]),
            duration_days=None,
            note="debt_core from Debt tag",
        )

    rec_non = _pick_first_instant_tag(
        df_all,
        end=end,
        tags=["LongTermDebtNoncurrent", "LongTermDebtAndCapitalLeaseObligations"],
        prefer_forms=prefer_forms,
    )
    rec_long = _pick_instant_tag(df_all, end=end, tag="LongTermDebt", prefer_forms=prefer_forms)
    rec_cur = _pick_first_instant_tag(
        df_all,
        end=end,
        tags=["LongTermDebtCurrent", "DebtCurrent"],
        prefer_forms=prefer_forms,
    )
    non_val = float(rec_non["val"]) if rec_non is not None and pd.notna(rec_non.get("val")) else None
    long_val = float(rec_long["val"]) if rec_long is not None and pd.notna(rec_long.get("val")) else None
    cur_val = float(rec_cur["val"]) if rec_cur is not None and pd.notna(rec_cur.get("val")) else None
    value, source_choice = choose_total_debt(non_val, cur_val, long_val)
    if value is None:
        return None

    used_recs: List[pd.Series] = []
    used_tags: List[str] = []
    note = ""
    source = "direct_component"
    if source_choice == "noncurrent_plus_current":
        if rec_non is not None:
            used_recs.append(rec_non)
            used_tags.append(str(rec_non.get("tag")))
        if rec_cur is not None:
            used_recs.append(rec_cur)
            used_tags.append(str(rec_cur.get("tag")))
        source = "derived_parts"
        note = "debt_core = noncurrent + current debt (preferred)"
    elif source_choice == "longterm_includes_current_or_total":
        if rec_long is not None:
            used_recs.append(rec_long)
            used_tags.append(str(rec_long.get("tag")))
        note = "debt_core from LongTermDebt (already includes current for some issuers)"
    elif source_choice == "noncurrent_only":
        if rec_non is not None:
            used_recs.append(rec_non)
            used_tags.append(str(rec_non.get("tag")))
        note = "debt_core from noncurrent debt only (current portion unavailable)"
    else:
        if rec_cur is not None:
            used_recs.append(rec_cur)
            used_tags.append(str(rec_cur.get("tag")))
        note = "debt_core fallback to current debt only"
    last_rec = _latest_rec(used_recs)

    return PickResult(
        value=float(value),
        source=source,
        source_choice=source_choice,
        tag=",".join(used_tags),
        accn=str(last_rec["accn"]) if last_rec is not None else None,
        form=str(last_rec["form"]) if last_rec is not None else None,
        filed=last_rec["filed_d"] if last_rec is not None else None,
        start=last_rec["start_d"] if last_rec is not None else None,
        end=last_rec["end_d"] if last_rec is not None else None,
        unit=str(last_rec["unit"]) if last_rec is not None else None,
        duration_days=None,
        note=note,
    )


def compute_lease_liabilities_instant(df_all: pd.DataFrame, end: dt.date, prefer_forms: List[str]) -> Optional[PickResult]:
    parts = []
    used_tags = []
    last_rec = None
    part_tags = [
        "OperatingLeaseLiabilityCurrent",
        "OperatingLeaseLiabilityNoncurrent",
        "FinanceLeaseLiabilityCurrent",
        "FinanceLeaseLiabilityNoncurrent",
        "LeaseLiabilityCurrent",
        "LeaseLiabilityNoncurrent",
    ]
    for tag in part_tags:
        rec = _pick_instant_tag(df_all, end=end, tag=tag, prefer_forms=prefer_forms)
        if rec is not None and pd.notna(rec["val"]):
            parts.append(float(rec["val"]))
            used_tags.append(tag)
            last_rec = rec
    if parts:
        note = "lease_liabilities = sum of lease liability components"
        return PickResult(
            value=float(sum(parts)),
            source="derived_parts",
            tag=",".join(used_tags),
            accn=str(last_rec["accn"]) if last_rec is not None else None,
            form=str(last_rec["form"]) if last_rec is not None else None,
            filed=last_rec["filed_d"] if last_rec is not None else None,
            start=last_rec["start_d"] if last_rec is not None else None,
            end=last_rec["end_d"] if last_rec is not None else None,
            unit=str(last_rec["unit"]) if last_rec is not None else None,
            duration_days=None,
            note=note,
        )
    # fallback to total tags if components not present
    total_tags = ["OperatingLeaseLiability", "FinanceLeaseLiability", "LeaseLiability", "LeaseLiabilities"]
    rec_total = _pick_first_instant_tag(df_all, end=end, tags=total_tags, prefer_forms=prefer_forms)
    if rec_total is None:
        return None
    return PickResult(
        value=float(rec_total["val"]),
        source="direct",
        tag=str(rec_total["tag"]) if "tag" in rec_total else "LeaseLiability",
        accn=str(rec_total["accn"]),
        form=str(rec_total["form"]),
        filed=rec_total["filed_d"],
        start=rec_total["start_d"],
        end=rec_total["end_d"],
        unit=str(rec_total["unit"]),
        duration_days=None,
        note="lease_liabilities from total lease tag",
    )


def compute_bank_deposits_instant(df_all: pd.DataFrame, end: dt.date, prefer_forms: List[str]) -> Optional[PickResult]:
    total_tags = [
        "CustomerDepositsAtPitneyBowesBank",
        "CustomerDeposits",
        "BankDeposits",
    ]
    current_tags = [
        "CustomerDepositsAtPitneyBowesBankCurrent",
        "CustomerDepositsCurrent",
    ]
    noncurrent_tags = [
        "CustomerDepositsAtPitneyBowesBankNoncurrent",
        "NoncurrentCustomerDepositsAtPitneyBowesBank",
        "CustomerDepositsNoncurrent",
    ]
    rec_cur = _pick_first_instant_tag(df_all, end=end, tags=current_tags, prefer_forms=prefer_forms)
    rec_non = _pick_first_instant_tag(df_all, end=end, tags=noncurrent_tags, prefer_forms=prefer_forms)
    if rec_cur is not None and rec_non is not None:
        val = float(rec_cur["val"]) + float(rec_non["val"])
        return PickResult(
            value=val,
            source="derived_parts",
            tag=f"{rec_cur['tag']},{rec_non['tag']}",
            accn=str(rec_cur["accn"]) if rec_cur is not None else None,
            form=str(rec_cur["form"]) if rec_cur is not None else None,
            filed=rec_cur["filed_d"] if rec_cur is not None else None,
            start=rec_cur["start_d"] if rec_cur is not None else None,
            end=rec_cur["end_d"] if rec_cur is not None else None,
            unit=str(rec_cur["unit"]) if rec_cur is not None else None,
            duration_days=None,
            note="bank_deposits = current + noncurrent",
        )
    rec_total = _pick_first_instant_tag(df_all, end=end, tags=total_tags, prefer_forms=prefer_forms)
    if rec_total is not None:
        return PickResult(
            value=float(rec_total["val"]),
            source="direct",
            tag=str(rec_total["tag"]),
            accn=str(rec_total["accn"]),
            form=str(rec_total["form"]),
            filed=rec_total["filed_d"],
            start=rec_total["start_d"],
            end=rec_total["end_d"],
            unit=str(rec_total["unit"]),
            duration_days=None,
            note="bank_deposits from total tag",
        )
    if rec_cur is not None:
        return PickResult(
            value=float(rec_cur["val"]),
            source="derived_parts",
            tag=str(rec_cur["tag"]),
            accn=str(rec_cur["accn"]),
            form=str(rec_cur["form"]),
            filed=rec_cur["filed_d"],
            start=rec_cur["start_d"],
            end=rec_cur["end_d"],
            unit=str(rec_cur["unit"]),
            duration_days=None,
            note="bank_deposits partial (current only)",
        )
    if rec_non is not None:
        return PickResult(
            value=float(rec_non["val"]),
            source="derived_parts",
            tag=str(rec_non["tag"]),
            accn=str(rec_non["accn"]),
            form=str(rec_non["form"]),
            filed=rec_non["filed_d"],
            start=rec_non["start_d"],
            end=rec_non["end_d"],
            unit=str(rec_non["unit"]),
            duration_days=None,
            note="bank_deposits partial (noncurrent only)",
        )
    return None


def compute_bank_finance_receivables_instant(df_all: pd.DataFrame, end: dt.date, prefer_forms: List[str]) -> Optional[PickResult]:
    total_tags = ["FinanceReceivablesNet", "FinanceReceivables", "FinanceReceivablesNetTotal"]
    short_tags = ["ShortTermFinanceReceivablesNet", "FinanceReceivablesCurrent", "FinanceReceivablesNetCurrent"]
    long_tags = ["LongTermFinanceReceivablesNet", "FinanceReceivablesNoncurrent", "FinanceReceivablesNetNoncurrent"]
    rec_short = _pick_first_instant_tag(df_all, end=end, tags=short_tags, prefer_forms=prefer_forms)
    rec_long = _pick_first_instant_tag(df_all, end=end, tags=long_tags, prefer_forms=prefer_forms)
    if rec_short is not None and rec_long is not None:
        val = float(rec_short["val"]) + float(rec_long["val"])
        return PickResult(
            value=val,
            source="derived_parts",
            tag=f"{rec_short['tag']},{rec_long['tag']}",
            accn=str(rec_short["accn"]) if rec_short is not None else None,
            form=str(rec_short["form"]) if rec_short is not None else None,
            filed=rec_short["filed_d"] if rec_short is not None else None,
            start=rec_short["start_d"] if rec_short is not None else None,
            end=rec_short["end_d"] if rec_short is not None else None,
            unit=str(rec_short["unit"]) if rec_short is not None else None,
            duration_days=None,
            note="bank_finance_receivables = short + long",
        )
    rec_total = _pick_first_instant_tag(df_all, end=end, tags=total_tags, prefer_forms=prefer_forms)
    if rec_total is not None:
        return PickResult(
            value=float(rec_total["val"]),
            source="direct",
            tag=str(rec_total["tag"]),
            accn=str(rec_total["accn"]),
            form=str(rec_total["form"]),
            filed=rec_total["filed_d"],
            start=rec_total["start_d"],
            end=rec_total["end_d"],
            unit=str(rec_total["unit"]),
            duration_days=None,
            note="bank_finance_receivables from total tag",
        )
    if rec_short is not None:
        return PickResult(
            value=float(rec_short["val"]),
            source="derived_parts",
            tag=str(rec_short["tag"]),
            accn=str(rec_short["accn"]),
            form=str(rec_short["form"]),
            filed=rec_short["filed_d"],
            start=rec_short["start_d"],
            end=rec_short["end_d"],
            unit=str(rec_short["unit"]),
            duration_days=None,
            note="bank_finance_receivables partial (short only)",
        )
    if rec_long is not None:
        return PickResult(
            value=float(rec_long["val"]),
            source="derived_parts",
            tag=str(rec_long["tag"]),
            accn=str(rec_long["accn"]),
            form=str(rec_long["form"]),
            filed=rec_long["filed_d"],
            start=rec_long["start_d"],
            end=rec_long["end_d"],
            unit=str(rec_long["unit"]),
            duration_days=None,
            note="bank_finance_receivables partial (long only)",
        )
    return None


def build_revolver_capacity_map(
    df_all: pd.DataFrame,
    hist: pd.DataFrame,
    prefer_forms: Optional[List[str]] = None,
) -> Tuple[Dict[pd.Timestamp, float], Dict[pd.Timestamp, Dict[str, Any]]]:
    """
    Build a per-quarter map for revolver borrowing capacity using XBRL instant facts.
    Tag: LineOfCreditFacilityMaximumBorrowingCapacity
    Chooses latest context end_d <= quarter_end. Returns (capacity_map, meta_map).
    """
    capacity_map: Dict[pd.Timestamp, float] = {}
    meta_map: Dict[pd.Timestamp, Dict[str, Any]] = {}
    if df_all is None or df_all.empty or hist is None or hist.empty:
        return capacity_map, meta_map

    tag = "LineOfCreditFacilityMaximumBorrowingCapacity"
    if "tag" not in df_all.columns:
        return capacity_map, meta_map
    s = df_all[(df_all["tag"] == tag) & df_all["end_d"].notna()].copy()
    if s.empty:
        return capacity_map, meta_map
    # USD instants only
    u = s["unit"].astype(str).str.upper()
    s = s[u.str.startswith("USD", na=False) & ~u.str.contains("/", na=False)].copy()
    if s.empty:
        return capacity_map, meta_map
    s["val"] = pd.to_numeric(s["val"], errors="coerce")
    s = s[s["val"].notna()]
    if s.empty:
        return capacity_map, meta_map

    if prefer_forms is None:
        prefer_forms = ["10-Q", "10-K"]

    # pick best record per end_d
    best_by_end: Dict[dt.date, pd.Series] = {}
    for end_d in sorted(s["end_d"].dropna().unique()):
        sub = s[s["end_d"] == end_d]
        rec = pick_best_instant(sub, end=end_d, prefer_forms=prefer_forms)
        if rec is not None and pd.notna(rec.get("val")):
            best_by_end[end_d] = rec

    if not best_by_end:
        return capacity_map, meta_map

    # map to each quarter: latest end_d <= quarter_end
    hq = pd.to_datetime(hist["quarter"], errors="coerce").dropna().unique()
    for q in sorted(hq):
        qd = pd.Timestamp(q).date()
        elig = [e for e in best_by_end.keys() if e <= qd]
        if not elig:
            continue
        end_d = max(elig)
        rec = best_by_end[end_d]
        capacity_map[pd.Timestamp(qd)] = float(rec["val"])
        meta_map[pd.Timestamp(qd)] = {
            "tag": tag,
            "accn": rec.get("accn"),
            "form": rec.get("form"),
            "filed": rec.get("filed_d"),
            "end_d": end_d,
        }
    return capacity_map, meta_map


DEBT_FAIR_VALUE_TAGS = [
    "LongTermDebtFairValue",
    "DebtFairValue",
    "DebtInstrumentFairValue",
    "LongTermDebtFairValueDisclosure",
]

_MONTH_MAP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def _normalize_tranche_name(name: str) -> str:
    s = str(name or "").strip().lower()
    s = re.sub(r"\bnotes\b", "note", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _clean_tranche_name(name: str) -> str:
    s = str(name or "").strip()
    if not s:
        return s
    # Keep instrument label up to explicit maturity phrase when present.
    m_due = re.search(
        r"(.+?\bdue\s+(?:in\s+)?(?:"
        r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|"
        r"sep(?:tember)?|sept(?:ember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?"
        r")\.?\s+\d{4})",
        s,
        flags=re.I,
    )
    if m_due:
        s = m_due.group(1)
    # If no month-based maturity, keep up to "due 20xx".
    m_due_y = re.search(r"(.+?\bdue\s+20\d{2})", s, flags=re.I)
    if m_due_y:
        s = m_due_y.group(1)
    # Remove trailing rate/amount fragments from OCR/table rows.
    s = re.sub(r"\s+(sofr|libor|prime|euribor)\b.*$", "", s, flags=re.I)
    s = re.sub(r"\s+\d+(?:\.\d+)?\s*%.*$", "", s, flags=re.I)
    s = re.sub(r"\s+\$?\d{1,3}(?:,\d{3})+(?:\.\d+)?(?:\s+\$?\d{1,3}(?:,\d{3})+(?:\.\d+)?)?.*$", "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" -,:;")
    return s


def _instrument_type_from_text(name: str, row_text: Optional[str] = None) -> str:
    txt = f"{name or ''} {row_text or ''}".lower()
    if "convertible" in txt:
        return "convertible"
    if "term loan" in txt:
        return "term_loan"
    if "revolver" in txt or "revolving credit" in txt:
        return "revolver"
    if "notes due" in txt or "senior notes" in txt or "note due" in txt:
        return "notes"
    return "other"


def _parse_maturity_from_text(name: str, row_text: Optional[str] = None) -> Tuple[Optional[dt.date], Optional[str], Optional[int]]:
    txt = f"{name or ''} {row_text or ''}"
    txt_low = txt.lower()
    m = re.search(
        r"(?:due\s+(?:in\s+)?)"
        r"(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|"
        r"sep(?:tember)?|sept(?:ember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\.?"
        r"\s+(\d{1,2})?,?\s*(20\d{2})",
        txt_low,
        re.I,
    )
    if m:
        mon_raw = str(m.group(1) or "").lower()
        mon = _MONTH_MAP.get(mon_raw, _MONTH_MAP.get(mon_raw[:3]))
        day = int(m.group(2)) if m.group(2) else 15
        year = int(m.group(3))
        if mon:
            try:
                dd = dt.date(year, mon, day)
            except Exception:
                dd = dt.date(year, mon, 15)
            month_name = dt.date(2000, mon, 1).strftime("%B")
            return dd, f"{month_name} {year}", year
    m2 = re.search(
        r"(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|"
        r"sep(?:tember)?|sept(?:ember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\.?\s+(20\d{2})",
        txt_low,
        re.I,
    )
    if m2:
        mon_raw = str(m2.group(1) or "").lower()
        mon = _MONTH_MAP.get(mon_raw, _MONTH_MAP.get(mon_raw[:3]))
        year = int(m2.group(2))
        if mon:
            dd = dt.date(year, mon, 15)
            month_name = dt.date(2000, mon, 1).strftime("%B")
            return dd, f"{month_name} {year}", year
    m3 = re.search(r"\b(20\d{2})\b", txt_low)
    if m3:
        year = int(m3.group(1))
        return None, None, year
    return None, None, None


def _tranche_doc_priority(source_kind: str, form: Optional[str]) -> int:
    sk = str(source_kind or "").lower()
    fm = str(form or "").upper()
    if sk == "slides_debt_profile":
        return 300
    if fm.startswith("10-K"):
        return 220
    if fm.startswith("10-Q"):
        return 200
    if fm.startswith("8-K"):
        return 160
    return 80


def _infer_tranche_meta(name: str, row_text: Optional[str] = None) -> Dict[str, Any]:
    nm = (name or "").strip()
    nml = nm.lower()
    rowt = (row_text or "").lower()
    maturity_date, maturity_display, year = _parse_maturity_from_text(nm, row_text)
    coupon = None
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", nml) or re.search(r"(\d+(?:\.\d+)?)\s*%", rowt)
    if m:
        try:
            coupon = float(m.group(1))
        except Exception:
            coupon = None
    spread = None
    m = re.search(r"(sofr|libor|prime|euribor)\s*\+\s*(\d+(?:\.\d+)?)\s*%", nml) or re.search(r"(sofr|libor|prime|euribor)\s*\+\s*(\d+(?:\.\d+)?)\s*%", rowt)
    if m:
        try:
            spread = float(m.group(2))
        except Exception:
            spread = None
    rate_type = _instrument_type_from_text(nm, row_text)
    if any(k in nml for k in ["sofr", "libor", "prime", "base rate", "floating", "variable"]) or any(
        k in rowt for k in ["sofr", "libor", "prime", "base rate", "floating", "variable"]
    ):
        rate_type = "float"
    elif rate_type == "other" and coupon is not None:
        rate_type = "fixed"
    if rate_type == "float":
        # For floating-rate tranches, treat percent as spread/margin
        if spread is None and coupon is not None:
            spread = coupon
        coupon = None
    return {
        "maturity_year": year,
        "maturity_date": maturity_date,
        "maturity_display": maturity_display,
        "instrument_type": _instrument_type_from_text(nm, row_text),
        "normalized_name": _normalize_tranche_name(nm),
        "coupon_pct": coupon,
        "spread_pct": spread,
        "rate_type": rate_type,
    }


def _tranche_family_key(name: str, instrument_type: str, amount_principal: float) -> str:
    core = _normalize_tranche_name(_clean_tranche_name(name))
    core = re.sub(
        r"\b(january|february|march|april|may|june|july|august|september|october|november|december|"
        r"jan|feb|mar|apr|jun|jul|aug|sep|sept|oct|nov|dec)\b",
        " ",
        core,
        flags=re.I,
    )
    core = re.sub(r"\b20\d{2}\b", " ", core)
    core = re.sub(r"\s+", " ", core).strip()
    amt_m = round(float(amount_principal) / 1e6, 3)
    return f"{instrument_type}|{core}|{amt_m:.3f}"


def _tranche_currency_from_text(name: str, row_text: Optional[str] = None) -> str:
    txt = f"{name or ''} {row_text or ''}".lower()
    if " eur" in txt or "euro" in txt or "€" in txt:
        return "eur"
    if " gbp" in txt or "sterling" in txt or "£" in txt:
        return "gbp"
    return "usd"


def _canonical_maturity_key(
    maturity_date: Optional[dt.date],
    maturity_year: Optional[int],
    maturity_display: Optional[str],
) -> str:
    if isinstance(maturity_date, dt.date):
        return maturity_date.strftime("%Y-%m")
    disp = str(maturity_display or "").strip().lower()
    if disp:
        m = re.search(
            r"\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|"
            r"aug(?:ust)?|sep(?:tember)?|sept(?:ember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b"
            r"[^\d]*(20\d{2})",
            disp,
            flags=re.I,
        )
        if m:
            mon_raw = str(m.group(1) or "").lower()
            mon = _MONTH_MAP.get(mon_raw, _MONTH_MAP.get(mon_raw[:3]))
            yr = int(m.group(2))
            if mon:
                return f"{yr:04d}-{int(mon):02d}"
        m2 = re.search(r"\b(20\d{2})\b", disp)
        if m2:
            return f"{int(m2.group(1)):04d}"
    if maturity_year is not None and pd.notna(maturity_year):
        try:
            return f"{int(maturity_year):04d}"
        except Exception:
            return "unknown"
    return "unknown"


def _tranche_canonical_key(
    clean_name: str,
    *,
    maturity_date: Optional[dt.date],
    maturity_year: Optional[int],
    maturity_display: Optional[str],
    instrument_type: str,
    row_text: Optional[str] = None,
) -> str:
    return (
        f"{_normalize_tranche_name(clean_name)}|"
        f"{_canonical_maturity_key(maturity_date, maturity_year, maturity_display)}|"
        f"{instrument_type}|"
        f"{_tranche_currency_from_text(clean_name, row_text)}"
    )


def _pick_debt_fair_value(df_all: pd.DataFrame, end: dt.date, prefer_forms: List[str]) -> Optional[PickResult]:
    tags = [t for t in DEBT_FAIR_VALUE_TAGS if t in set(df_all.get("tag", []))]
    if not tags:
        tags = [t for t in df_all.get("tag", []).dropna().unique() if "Debt" in str(t) and "FairValue" in str(t)]
    for tag in tags:
        rec = _pick_instant_tag(df_all, end=end, tag=tag, prefer_forms=prefer_forms)
        if rec is not None and pd.notna(rec.get("val")):
            return PickResult(
                value=float(rec["val"]),
                source="direct",
                tag=str(tag),
                accn=str(rec["accn"]),
                form=str(rec["form"]),
                filed=rec["filed_d"],
                start=rec["start_d"],
                end=rec["end_d"],
                unit=str(rec["unit"]),
                duration_days=None,
                note=f"debt fair value from {tag}",
            )
    return None


def build_debt_profile(
    hist: pd.DataFrame,
    df_all: pd.DataFrame,
    debt_tranches: pd.DataFrame,
    slides_debt: Optional[pd.DataFrame] = None,
    debt_schedule: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if hist is None or hist.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    h = hist.copy()
    h["quarter"] = pd.to_datetime(h["quarter"], errors="coerce")
    h = h[h["quarter"].notna()]
    if h.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    latest_q = h["quarter"].max()
    latest_row = h[h["quarter"] == latest_q].iloc[0]
    as_of_date = latest_q.date()

    def _get_hist(col: str) -> Optional[float]:
        if col not in h.columns:
            return None
        v = pd.to_numeric(latest_row.get(col), errors="coerce")
        return float(v) if pd.notna(v) else None

    prefer_forms = ["10-Q", "10-K"]
    debt_current = None
    debt_long = None
    pr_cur = _pick_instant_tag(df_all, end=latest_q.date(), tag="LongTermDebtCurrent", prefer_forms=prefer_forms)
    if pr_cur is None:
        pr_cur = _pick_instant_tag(df_all, end=latest_q.date(), tag="DebtCurrent", prefer_forms=prefer_forms)
    if pr_cur is not None and pd.notna(pr_cur.get("val")):
        debt_current = float(pr_cur["val"])
    pr_lt = _pick_instant_tag(df_all, end=latest_q.date(), tag="LongTermDebt", prefer_forms=prefer_forms)
    if pr_lt is None:
        pr_lt = _pick_instant_tag(df_all, end=latest_q.date(), tag="LongTermDebtAndCapitalLeaseObligations", prefer_forms=prefer_forms)
    if pr_lt is not None and pd.notna(pr_lt.get("val")):
        debt_long = float(pr_lt["val"])
    pr_fv = _pick_debt_fair_value(df_all, end=latest_q.date(), prefer_forms=prefer_forms)
    debt_fair = float(pr_fv.value) if pr_fv is not None and pr_fv.value is not None else None

    rows = []

    def _add(metric: str, value: Optional[float], note: str, source: str) -> None:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return
        rows.append({
            "quarter": latest_q.date(),
            "metric": metric,
            "value": float(value),
            "source": source,
            "note": note,
        })

    _add("total_debt", _get_hist("total_debt"), "History_Q total_debt (legacy).", "History_Q")
    _add("debt_core", _get_hist("debt_core"), "Corporate borrowings (long-term + current portion).", "History_Q")
    _add("debt_current", debt_current, "Current portion of long-term debt.", "XBRL")
    _add("debt_long_term", debt_long, "Long-term debt (carrying value).", "XBRL")
    _add("debt_fair_value", debt_fair, f"Debt fair value from XBRL ({pr_fv.tag})" if pr_fv else "Debt fair value from XBRL.", "XBRL")
    _add("lease_liabilities", _get_hist("lease_liabilities"), "Operating + finance leases.", "History_Q")
    _add("bank_deposits", _get_hist("bank_deposits"), "Customer deposits (bank).", "History_Q")
    _add("bank_finance_receivables", _get_hist("bank_finance_receivables"), "Finance receivables (short + long).", "History_Q")
    _add("bank_net_funding", _get_hist("bank_net_funding"), "Deposits - finance receivables.", "History_Q")
    _add("cash", _get_hist("cash"), "Cash & equivalents.", "History_Q")
    if _get_hist("debt_core") is not None and _get_hist("cash") is not None:
        _add("net_debt_core", _get_hist("debt_core") - _get_hist("cash"), "Debt core - cash.", "Derived")

    profile_df = pd.DataFrame(rows)

    tr_latest = pd.DataFrame()
    maturity_df = pd.DataFrame()
    qa_rows: List[Dict[str, Any]] = []
    info_rows: List[Dict[str, Any]] = []

    def _as_of_gap_days(v: Any) -> int:
        try:
            d = pd.to_datetime(v, errors="coerce")
            if pd.isna(d):
                return 9999
            return abs((d.date() - as_of_date).days)
        except Exception:
            return 9999

    candidates: List[Dict[str, Any]] = []
    slide_principal_total: Optional[float] = None

    if slides_debt is not None and not slides_debt.empty:
        sd = slides_debt.copy()
        if "quarter" in sd.columns:
            sd["quarter"] = pd.to_datetime(sd["quarter"], errors="coerce")
            sd = sd[sd["quarter"].notna()]
            sd = sd[sd["quarter"].dt.to_period("Q") == latest_q.to_period("Q")]
        for _, r in sd.iterrows():
            raw_name = str(r.get("tranche") or r.get("tranche_name") or "").strip()
            if not raw_name:
                continue
            clean_name = _clean_tranche_name(raw_name) or raw_name
            amt = pd.to_numeric(r.get("amount"), errors="coerce")
            if "principal amount" in raw_name.lower():
                if pd.notna(amt) and float(amt) > 0:
                    slide_principal_total = float(amt)
                continue
            if pd.isna(amt) or float(amt) <= 0:
                continue
            meta = _infer_tranche_meta(clean_name, raw_name)
            instrument_type = str(meta.get("instrument_type") or _instrument_type_from_text(clean_name, raw_name))
            maturity_date = meta.get("maturity_date")
            maturity_year = meta.get("maturity_year")
            maturity_display = meta.get("maturity_display")
            tranche_key = _tranche_canonical_key(
                clean_name,
                maturity_date=maturity_date,
                maturity_year=int(maturity_year) if pd.notna(maturity_year) else None,
                maturity_display=str(maturity_display or ""),
                instrument_type=instrument_type,
                row_text=raw_name,
            )
            family_key = _tranche_family_key(clean_name, instrument_type, float(amt))
            candidates.append(
                {
                    "quarter": latest_q,
                    "tranche_name": clean_name,
                    "amount_principal": float(amt),
                    "amount_carrying": None,
                    "amount": float(amt),
                    "row_text": raw_name,
                    "maturity_year": int(maturity_year) if pd.notna(maturity_year) else None,
                    "maturity_date": maturity_date,
                    "maturity_display": maturity_display,
                    "instrument_type": instrument_type,
                    "coupon_pct": meta.get("coupon_pct"),
                    "spread_pct": meta.get("spread_pct"),
                    "rate_type": meta.get("rate_type"),
                    "source_kind": "slides_debt_profile",
                    "source_priority": _tranche_doc_priority("slides_debt_profile", None),
                    "parse_quality": "asof_matched",
                    "parse_quality_score": 2,
                    "period_match_score": 1,
                    "form": "slides",
                    "accn": None,
                    "doc": r.get("doc"),
                    "filed": None,
                    "report_date": pd.to_datetime(r.get("quarter"), errors="coerce").date() if pd.notna(pd.to_datetime(r.get("quarter"), errors="coerce")) else None,
                    "asof_gap_days": _as_of_gap_days(r.get("quarter")),
                    "tranche_key": tranche_key,
                    "family_key": family_key,
                }
            )

    if debt_tranches is not None and not debt_tranches.empty:
        dtl = debt_tranches.copy()
        dtl["quarter"] = pd.to_datetime(dtl["quarter"], errors="coerce")
        dtl = dtl[dtl["quarter"].notna()]
        dtl = dtl[dtl["quarter"].dt.to_period("Q") == latest_q.to_period("Q")]
        for _, r in dtl.iterrows():
            raw_name = str(r.get("tranche_name") or "").strip()
            if not raw_name:
                continue
            clean_name = _clean_tranche_name(raw_name) or raw_name
            amt = pd.to_numeric(r.get("amount"), errors="coerce")
            if pd.isna(amt) or float(amt) <= 0:
                continue
            row_text = str(r.get("row_text") or "")
            meta = _infer_tranche_meta(clean_name, row_text)
            instrument_type = str(meta.get("instrument_type") or _instrument_type_from_text(clean_name, row_text))
            maturity_date = meta.get("maturity_date")
            maturity_year = meta.get("maturity_year")
            maturity_display = meta.get("maturity_display")
            tranche_key = _tranche_canonical_key(
                clean_name,
                maturity_date=maturity_date,
                maturity_year=int(maturity_year) if pd.notna(maturity_year) else None,
                maturity_display=str(maturity_display or ""),
                instrument_type=instrument_type,
                row_text=row_text,
            )
            family_key = _tranche_family_key(clean_name, instrument_type, float(amt))
            form = str(r.get("form") or "")
            parse_quality = str(r.get("parse_quality") or "")
            parse_quality_score = 2 if parse_quality == "asof_matched" else (1 if parse_quality else 0)
            candidates.append(
                {
                    "quarter": latest_q,
                    "tranche_name": clean_name,
                    "amount_principal": float(amt),
                    "amount_carrying": None,
                    "amount": float(amt),
                    "row_text": row_text,
                    "maturity_year": int(maturity_year) if pd.notna(maturity_year) else None,
                    "maturity_date": maturity_date,
                    "maturity_display": maturity_display,
                    "instrument_type": instrument_type,
                    "coupon_pct": meta.get("coupon_pct"),
                    "spread_pct": meta.get("spread_pct"),
                    "rate_type": meta.get("rate_type"),
                    "source_kind": "filing_debt_table",
                    "source_priority": _tranche_doc_priority("filing_debt_table", form),
                    "parse_quality": parse_quality,
                    "parse_quality_score": parse_quality_score,
                    "form": form,
                    "accn": r.get("accn"),
                    "doc": r.get("doc"),
                    "filed": pd.to_datetime(r.get("filed"), errors="coerce"),
                    "report_date": pd.to_datetime(r.get("report_date"), errors="coerce"),
                    "table_total_debt": r.get("table_total_debt"),
                    "table_total_long_term_debt": r.get("table_total_long_term_debt"),
                    "table_total_label": r.get("table_total_label"),
                    "scale_applied": r.get("scale_applied"),
                    "period_match": r.get("period_match"),
                    "period_match_score": 1 if bool(r.get("period_match")) else 0,
                    "asof_col_date": pd.to_datetime(r.get("asof_col_date"), errors="coerce"),
                    "asof_select_method": r.get("asof_select_method"),
                    "amount_col_idx": r.get("amount_col_idx"),
                    "asof_gap_days": _as_of_gap_days(r.get("report_date") or r.get("quarter")),
                    "tranche_key": tranche_key,
                    "family_key": family_key,
                }
            )

    duplicate_tranche_count = 0
    family_conflict_drop_count = 0
    bad_maturity_parse_count = 0
    if candidates:
        cdf = pd.DataFrame(candidates)
        cdf["filed"] = pd.to_datetime(cdf.get("filed"), errors="coerce")
        if "period_match_score" not in cdf.columns:
            cdf["period_match_score"] = 0
        else:
            cdf["period_match_score"] = pd.to_numeric(cdf.get("period_match_score"), errors="coerce").fillna(0)
        if "parse_quality_score" not in cdf.columns:
            cdf["parse_quality_score"] = cdf.get("parse_quality").astype(str).str.lower().map(
                {"asof_matched": 2, "fallback_amount_col": 0}
            ).fillna(1)
        else:
            cdf["parse_quality_score"] = pd.to_numeric(cdf.get("parse_quality_score"), errors="coerce").fillna(0)
        cdf["amount_nonzero"] = pd.to_numeric(cdf.get("amount_principal"), errors="coerce").fillna(0).abs() > 0
        cdf["maturity_quality"] = cdf.get("maturity_date").apply(lambda d: 1 if isinstance(d, dt.date) else 0)
        has_slide_rows = bool((cdf.get("source_kind").astype(str) == "slides_debt_profile").any())
        has_filing_asof_match = bool(
            (
                (cdf.get("source_kind").astype(str) == "filing_debt_table")
                & (cdf["period_match_score"] > 0)
            ).any()
        )
        if has_slide_rows or has_filing_asof_match:
            stale_filing_mask = (
                (cdf.get("source_kind").astype(str) == "filing_debt_table")
                & (cdf["period_match_score"] <= 0)
                & (cdf["parse_quality_score"] <= 0)
            )
            stale_drop_n = int(stale_filing_mask.sum())
            if stale_drop_n > 0:
                cdf = cdf.loc[~stale_filing_mask].copy()
                info_rows.append(
                    {
                        "quarter": as_of_date,
                        "metric": "debt_profile_asof_mismatch_filtered",
                        "severity": "info",
                        "message": (
                            f"Dropped {stale_drop_n} filing tranche rows where amount column did not match latest "
                            f"as-of period for {as_of_date}."
                        ),
                        "source": "Debt_Tranches_Latest",
                    }
                )
        cdf = cdf.sort_values(
            [
                "source_priority",
                "period_match_score",
                "parse_quality_score",
                "amount_nonzero",
                "maturity_quality",
                "asof_gap_days",
                "filed",
            ],
            ascending=[False, False, False, False, False, True, False],
        ).reset_index(drop=True)
        duplicate_tranche_count = int(len(cdf) - len(cdf.drop_duplicates(subset=["tranche_key"])))
        cdf = cdf.drop_duplicates(subset=["tranche_key"], keep="first").copy()
        cdf = cdf.sort_values(
            [
                "source_priority",
                "period_match_score",
                "parse_quality_score",
                "amount_nonzero",
                "maturity_quality",
                "asof_gap_days",
                "filed",
            ],
            ascending=[False, False, False, False, False, True, False],
        )
        amt_all = pd.to_numeric(cdf.get("amount_principal"), errors="coerce").dropna()
        if not amt_all.empty:
            med_amt = float(amt_all.median())
            tiny_floor = max(5_000_000.0, 0.01 * max(1.0, med_amt))
            tiny_mask = (
                (pd.to_numeric(cdf.get("amount_principal"), errors="coerce").fillna(0).abs() < tiny_floor)
                & (cdf.get("source_kind").astype(str) == "filing_debt_table")
            )
            tiny_drop_n = int(tiny_mask.sum())
            if tiny_drop_n > 0:
                cdf = cdf.loc[~tiny_mask].copy()
                info_rows.append(
                    {
                        "quarter": as_of_date,
                        "metric": "debt_profile_tiny_amount_filtered",
                        "severity": "info",
                        "message": f"Filtered {tiny_drop_n} suspicious tiny filing debt rows (<${tiny_floor/1e6:,.2f}m).",
                        "source": "Debt_Tranches_Latest",
                    }
                )
        kept_idx: List[int] = []
        for _, grp in cdf.groupby("family_key", sort=False):
            if len(grp) <= 1:
                kept_idx.extend(grp.index.tolist())
                continue
            grp2 = grp.sort_values(
                [
                    "source_priority",
                    "period_match_score",
                    "parse_quality_score",
                    "amount_nonzero",
                    "maturity_quality",
                    "asof_gap_days",
                    "filed",
                ],
                ascending=[False, False, False, False, False, True, False],
            )
            kept_idx.append(int(grp2.index[0]))
            family_conflict_drop_count += int(len(grp2) - 1)
        cdf = cdf.loc[kept_idx].copy()
        cdf["maturity_date_sort"] = pd.to_datetime(cdf.get("maturity_date"), errors="coerce")
        cdf = cdf.sort_values(
            ["maturity_date_sort", "maturity_year", "tranche_name"],
            ascending=[True, True, True],
            na_position="last",
        ).reset_index(drop=True)
        cdf = cdf.drop(columns=["maturity_date_sort"], errors="ignore")

        def _near_term(md: Any) -> bool:
            if not isinstance(md, dt.date):
                return False
            dd = (md - as_of_date).days
            return bool(0 <= dd <= 730)

        cdf["near_term"] = cdf.get("maturity_date").apply(_near_term)
        cdf["quarter"] = pd.to_datetime(cdf["quarter"], errors="coerce").dt.date
        bad_maturity_parse_count = int((~cdf["maturity_year"].notna()).sum())
        tr_latest = cdf.copy()

        if "maturity_year" in tr_latest.columns:
            maturity_df = (
                tr_latest.dropna(subset=["maturity_year"])
                .groupby("maturity_year", as_index=False)["amount_principal"]
                .sum()
                .rename(columns={"amount_principal": "amount_total"})
            )
            if not maturity_df.empty:
                maturity_df["quarter"] = latest_q.date()
                maturity_df["maturity_label"] = maturity_df["maturity_year"].astype(int).astype(str)
                maturity_df["source_kind"] = "Debt_Tranches_Latest"
                maturity_df["source_basis"] = "principal_tranche_sum"

        # Weighted average coupon/spread if available (principal-weighted).
        try:
            fixed = tr_latest.dropna(subset=["coupon_pct", "amount_principal"])
            if not fixed.empty:
                wac = (fixed["coupon_pct"] * fixed["amount_principal"]).sum() / fixed["amount_principal"].sum()
                profile_df = pd.concat([profile_df, pd.DataFrame([{
                    "quarter": latest_q.date(),
                    "metric": "weighted_avg_coupon_pct",
                    "value": float(wac),
                    "source": "Debt_Tranches_Latest",
                    "note": "Weighted avg coupon (fixed-rate tranches).",
                }])], ignore_index=True)
            flt = tr_latest.dropna(subset=["spread_pct", "amount_principal"])
            if not flt.empty:
                was = (flt["spread_pct"] * flt["amount_principal"]).sum() / flt["amount_principal"].sum()
                profile_df = pd.concat([profile_df, pd.DataFrame([{
                    "quarter": latest_q.date(),
                    "metric": "weighted_avg_spread_pct",
                    "value": float(was),
                    "source": "Debt_Tranches_Latest",
                    "note": "Weighted avg spread (floating tranches).",
                }])], ignore_index=True)
        except Exception:
            pass

    principal_total = float(pd.to_numeric(tr_latest.get("amount_principal"), errors="coerce").dropna().sum()) if not tr_latest.empty else None
    carrying_total = _get_hist("debt_core")
    bs_total_debt = _get_hist("total_debt")
    long_term_book_value = debt_long
    near_term_total = float(pd.to_numeric(tr_latest.loc[tr_latest["near_term"] == True, "amount_principal"], errors="coerce").dropna().sum()) if not tr_latest.empty else 0.0

    schedule_latest = pd.DataFrame()
    schedule_total = None
    if debt_schedule is not None and not debt_schedule.empty:
        schedule_latest = debt_schedule.copy()
        schedule_latest["quarter"] = pd.to_datetime(schedule_latest.get("quarter"), errors="coerce")
        schedule_latest = schedule_latest[schedule_latest["quarter"].notna()]
        schedule_latest = schedule_latest[schedule_latest["quarter"].dt.to_period("Q") == latest_q.to_period("Q")]
        if not schedule_latest.empty:
            schedule_total = float(pd.to_numeric(schedule_latest.get("amount_total"), errors="coerce").dropna().sum())

    selected_basis_value = long_term_book_value if long_term_book_value is not None else carrying_total
    selected_basis_metric = "debt_long_term" if long_term_book_value is not None else "debt_core"
    tieout_diff_pct = None
    schedule_diff_pct = None
    if principal_total is not None and selected_basis_value not in (None, 0):
        tieout_diff_pct = abs(float(principal_total) - float(selected_basis_value)) / abs(float(selected_basis_value))
    if schedule_total is not None and selected_basis_value not in (None, 0):
        schedule_diff_pct = abs(float(schedule_total) - float(selected_basis_value)) / abs(float(selected_basis_value))

    needs_tranche_review = bool(tieout_diff_pct is not None and tieout_diff_pct > 0.02)
    if needs_tranche_review:
        qa_rows.append({
            "quarter": as_of_date,
            "metric": "debt_latest_publish_guardrail",
            "check": "debt_latest_publish_guardrail",
            "status": "fail",
            "value": tieout_diff_pct,
            "message": (
                f"Latest tranche tie-out failed vs {selected_basis_metric}: "
                f"{float(principal_total or 0.0)/1e6:,.3f}m vs {float(selected_basis_value or 0.0)/1e6:,.3f}m "
                f"(diff {float(tieout_diff_pct or 0.0):.2%})."
            ),
        })
        info_rows.append({
            "quarter": as_of_date,
            "metric": "debt_latest_publish_guardrail",
            "severity": "warn",
            "message": (
                f"Suppressing tranche-level latest debt publishing because tie-out failed vs {selected_basis_metric} "
                f"by {float(tieout_diff_pct or 0.0):.2%}."
            ),
            "source": "Debt_Tranches_Latest",
        })
        if not schedule_latest.empty:
            maturity_df = schedule_latest[["quarter", "maturity_year", "maturity_label", "amount_total", "source_kind"]].copy()
            maturity_df["quarter"] = pd.to_datetime(maturity_df["quarter"], errors="coerce").dt.date
            maturity_df["source_basis"] = "principal_excl_issuance_costs"
            principal_total = schedule_total
            near_term_total = float(
                pd.to_numeric(
                    maturity_df.loc[pd.to_numeric(maturity_df["maturity_year"], errors="coerce").fillna(9999) <= 2027, "amount_total"],
                    errors="coerce",
                ).dropna().sum()
            )
            info_rows.append({
                "quarter": as_of_date,
                "metric": "debt_schedule_fallback_used",
                "severity": "info" if schedule_diff_pct is not None and schedule_diff_pct <= 0.02 else "warn",
                "message": (
                    f"Used scheduled repayments fallback for maturity outputs; total {float(schedule_total or 0.0)/1e6:,.3f}m "
                    f"vs {selected_basis_metric} diff {float(schedule_diff_pct or 0.0):.2%}."
                ),
                "source": "Debt_Maturity_Ladder",
            })
        else:
            maturity_df = pd.DataFrame([{
                "quarter": as_of_date,
                "maturity_year": None,
                "maturity_label": "Needs review: tranche tie-out failed",
                "amount_total": None,
                "source_kind": "qa_guardrail",
                "source_basis": "",
            }])
        tr_latest = pd.DataFrame([{
            "quarter": as_of_date,
            "tranche_name": "Needs review: tranche tie-out failed",
            "amount_principal": None,
            "amount_carrying": None,
            "maturity_display": None,
            "rate_type": None,
            "coupon_pct": None,
            "spread_pct": None,
            "near_term": None,
            "source_kind": "qa_guardrail",
            "qa_status": "FAIL",
            "review_note": (
                f"Suppressed latest tranche publish because principal tie-out vs {selected_basis_metric} exceeded 2%."
            ),
        }])

    carrying_minus_principal = (
        (float(carrying_total) - float(principal_total))
        if carrying_total is not None and principal_total is not None
        else None
    )

    if principal_total is not None:
        _add("debt_principal_total", principal_total, "Sum of latest published debt principal basis.", "Debt_Tranches_Latest")
    if carrying_total is not None:
        _add("debt_carrying_total", carrying_total, "Debt carrying total (latest, History_Q debt_core).", "History_Q")
    if carrying_minus_principal is not None:
        _add("debt_net_discounts_issuance_costs", carrying_minus_principal, "Carrying total minus latest published principal total.", "Derived")
    if slide_principal_total is not None:
        _add("debt_slide_principal_total", slide_principal_total, "Principal total disclosed in debt profile slide table.", "slides_debt_profile")

    def _qa_status(diff_abs: float, warn_thr: float, fail_thr: float) -> str:
        if diff_abs <= warn_thr:
            return "pass"
        if diff_abs <= fail_thr:
            return "warn"
        return "fail"

    # QA: principal sum vs slide principal total.
    if principal_total is not None and slide_principal_total is not None:
        diff = float(principal_total) - float(slide_principal_total)
        status = _qa_status(abs(diff), 1_000_000.0, 5_000_000.0)
        qa_rows.append({
            "quarter": as_of_date,
            "metric": "debt_profile_principal_sum_vs_slide_total",
            "check": "debt_profile_principal_sum_vs_slide_total",
            "status": status,
            "value": diff,
            "message": f"Principal sum {principal_total/1e6:,.3f}m vs slide total {slide_principal_total/1e6:,.3f}m (diff {diff/1e6:,.3f}m).",
        })
        info_rows.append({
            "quarter": as_of_date,
            "metric": "debt_profile_principal_sum_vs_slide_total",
            "severity": "info" if status == "pass" else "warn",
            "message": f"Principal sum {principal_total/1e6:,.3f}m vs slide total {slide_principal_total/1e6:,.3f}m.",
            "source": "Debt_Tranches_Latest/Slides_Debt_Profile",
        })

    # QA: carrying vs BS debt.
    if carrying_total is not None and bs_total_debt is not None:
        diff = float(carrying_total) - float(bs_total_debt)
        status = _qa_status(abs(diff), 5_000_000.0, 25_000_000.0)
        qa_rows.append({
            "quarter": as_of_date,
            "metric": "debt_core_carrying_vs_bs_debt",
            "check": "debt_core_carrying_vs_bs_debt",
            "status": status,
            "value": diff,
            "message": f"Carrying debt_core {carrying_total/1e6:,.3f}m vs BS total_debt {bs_total_debt/1e6:,.3f}m (diff {diff/1e6:,.3f}m).",
        })
        info_rows.append({
            "quarter": as_of_date,
            "metric": "debt_core_carrying_vs_bs_debt",
            "severity": "info" if status == "pass" else "warn",
            "message": f"debt_core {carrying_total/1e6:,.3f}m vs total_debt {bs_total_debt/1e6:,.3f}m.",
            "source": "History_Q",
        })

    # QA: duplicates and maturity parsing.
    post_dedup_dup = 0
    if not tr_latest.empty and "tranche_key" in tr_latest.columns:
        post_dedup_dup = int(len(tr_latest) - len(tr_latest.drop_duplicates(subset=["tranche_key"])))
    qa_rows.append({
        "quarter": as_of_date,
        "metric": "duplicate_tranche_count",
        "check": "duplicate_tranche_count",
        "status": "pass" if post_dedup_dup == 0 else "fail",
        "value": int(post_dedup_dup),
        "message": f"Post-dedup duplicate tranche rows: {post_dedup_dup}.",
    })
    if (duplicate_tranche_count + family_conflict_drop_count) > 0:
        info_rows.append(
            {
                "quarter": as_of_date,
                "metric": "duplicate_tranche_count",
                "severity": "info",
                "message": (
                    f"Collapsed source duplicates before final latest view: "
                    f"exact={duplicate_tranche_count}, family_conflict={family_conflict_drop_count}."
                ),
                "source": "Debt_Tranches_Latest",
            }
        )
    qa_rows.append({
        "quarter": as_of_date,
        "metric": "bad_maturity_parse_count",
        "check": "bad_maturity_parse_count",
        "status": "pass" if bad_maturity_parse_count == 0 else "warn",
        "value": int(bad_maturity_parse_count),
        "message": f"Tranches with unresolved maturity parse: {bad_maturity_parse_count}.",
    })
    if principal_total is not None:
        near_status = "pass" if near_term_total <= float(principal_total) + 1.0 else "fail"
        qa_rows.append({
            "quarter": as_of_date,
            "metric": "near_term_maturities_sanity",
            "check": "near_term_maturities_sanity",
            "status": near_status,
            "value": near_term_total - float(principal_total),
            "message": f"Near-term principal {near_term_total/1e6:,.3f}m vs total principal {principal_total/1e6:,.3f}m.",
        })

    qa_df = pd.DataFrame(qa_rows)
    info_df = pd.DataFrame(info_rows)
    return profile_df, tr_latest, maturity_df, qa_df, info_df


def build_debt_credit_notes(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_docs: int = 8,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    patterns = {
        "revolver": re.compile(r"revolver|revolving credit facility|credit facility", re.I),
        "covenant": re.compile(r"covenant|interest coverage|leverage ratio|secured net leverage|total net leverage", re.I),
        "ratings": re.compile(r"moody|standard\\s*&\\s*poor|s&p|fitch|rating", re.I),
        "maturity": re.compile(r"springing maturity|maturity", re.I),
    }

    def _snippet(txt: str, start: int, end: int, window: int = 140) -> str:
        s = max(0, start - window)
        e = min(len(txt), end + window)
        snip = txt[s:e]
        return re.sub(r"\\s+", " ", snip).strip()

    scanned = 0
    for batch in _iter_submission_batches(sec, submissions):
        forms = batch.get("form", [])
        accns = batch.get("accessionNumber", [])
        report_dates = batch.get("reportDate", []) or []
        filing_dates = batch.get("filingDate", []) or []
        primary_docs = batch.get("primaryDocument", []) or []
        n = min(len(forms), len(accns))
        for i in range(n):
            form = forms[i]
            if form not in ("10-Q", "10-Q/A", "10-K", "10-K/A", "8-K", "8-K/A"):
                continue
            accn = accns[i]
            doc = primary_docs[i] if i < len(primary_docs) else None
            if not doc:
                continue
            rep = report_dates[i] if i < len(report_dates) else None
            fdate = filing_dates[i] if i < len(filing_dates) else None
            q_end = parse_date(rep) or parse_date(fdate)
            if not _is_quarter_end(q_end):
                q_end = _coerce_prev_quarter_end(q_end)
            accn_nd = normalize_accession(accn)
            try:
                html_bytes = sec.download_document(cik_int, accn_nd, doc)
            except Exception:
                continue
            text = strip_html(html_bytes.decode("utf-8", errors="ignore"))
            if not text:
                continue
            for cat, rx in patterns.items():
                hits = 0
                for m in rx.finditer(text):
                    rows.append({
                        "quarter": q_end,
                        "accn": accn,
                        "form": form,
                        "filed": parse_date(fdate),
                        "report_date": parse_date(rep),
                        "doc": doc,
                        "category": cat,
                        "snippet": _snippet(text, m.start(), m.end()),
                        "source_class": "filing_text",
                        "method": "debt_text_scan",
                        "qa_severity": "warn",
                    })
                    hits += 1
                    if hits >= 3:
                        break
            scanned += 1
            if scanned >= max_docs:
                break
        if scanned >= max_docs:
            break

    return pd.DataFrame(rows)


def _build_company_overview_legacy(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    *,
    ticker: Optional[str] = None,
) -> Dict[str, Any]:
    """Build deterministic company overview text + revenue stream mix from latest 10-K."""
    out: Dict[str, Any] = {
        "what_it_does": "N/A",
        "what_it_does_source": "Source: N/A (latest 10-K not found)",
        "key_advantage": "N/A",
        "key_advantage_source": "Source: N/A (latest 10-K not found)",
        "segment_operating_model": [],
        "segment_operating_model_source": "Source: N/A (segment operating model not found)",
        "key_dependencies": [],
        "key_dependencies_source": "Source: N/A (Item 1A dependencies not found)",
        "wrong_thesis_bullets": [],
        "wrong_thesis_source": "Source: N/A (Item 1A wrong-thesis bullets not found)",
        "revenue_streams": [],
        "revenue_streams_source": "Source: N/A (latest 10-K revenue stream table not parsed)",
        "asof_fy_end": None,
        "revenue_streams_period": None,
    }

    tenk_rows: List[Dict[str, Any]] = []
    for batch in _iter_submission_batches(sec, submissions):
        forms = batch.get("form", []) or []
        accns = batch.get("accessionNumber", []) or []
        report_dates = batch.get("reportDate", []) or []
        filing_dates = batch.get("filingDate", []) or []
        primary_docs = batch.get("primaryDocument", []) or []
        n = min(len(forms), len(accns))
        for i in range(n):
            form = str(forms[i] or "").upper().strip()
            if not form.startswith("10-K"):
                continue
            accn = str(accns[i] or "").strip()
            doc = str(primary_docs[i] or "") if i < len(primary_docs) else ""
            if not accn or not doc:
                continue
            filed = parse_date(filing_dates[i]) if i < len(filing_dates) else None
            report = parse_date(report_dates[i]) if i < len(report_dates) else None
            tenk_rows.append(
                {
                    "form": form,
                    "accn": accn,
                    "doc": doc,
                    "filed": filed,
                    "report": report,
                }
            )

    if not tenk_rows:
        return out

    tenk_rows = sorted(
        tenk_rows,
        key=lambda r: (
            pd.Timestamp(r.get("filed")) if r.get("filed") is not None else pd.Timestamp("1900-01-01"),
            pd.Timestamp(r.get("report")) if r.get("report") is not None else pd.Timestamp("1900-01-01"),
        ),
        reverse=True,
    )
    sel = tenk_rows[0]
    accn = str(sel.get("accn") or "")
    doc = str(sel.get("doc") or "")
    form = str(sel.get("form") or "10-K")
    filed = sel.get("filed")
    report = sel.get("report")
    out["asof_fy_end"] = report

    accn_nd = normalize_accession(accn)
    try:
        html_bytes = sec.download_document(cik_int, accn_nd, doc)
    except Exception:
        src = f"Source: SEC {form} accn={accn} filed={filed or 'n/a'} (primary doc download failed)"
        out["what_it_does_source"] = src
        out["key_advantage_source"] = src
        out["revenue_streams_source"] = src
        return out

    html_text = html_bytes.decode("utf-8", errors="ignore")
    plain = re.sub(r"\s+", " ", strip_html(html_text)).strip()
    low = plain.lower()

    # Item 1 / Item 1A extraction (skip table-of-contents hits when possible).
    item1_matches = list(re.finditer(r"\bitem\s*1\.?\s*business\b", low, re.I))
    item1a_matches = list(re.finditer(r"\bitem\s*1a\.?\s*risk\s+factors?\b|\bitem\s*1a\.?\b", low, re.I))
    item1b_matches = list(re.finditer(r"\bitem\s*1b\.?\b", low, re.I))
    item2_matches = list(re.finditer(r"\bitem\s*2\.?\b", low, re.I))

    def _pick_non_toc(matches: List[re.Match], *, min_pos: int = 5000) -> Optional[re.Match]:
        if not matches:
            return None
        later = [m for m in matches if int(m.start()) >= int(min_pos)]
        if later:
            return later[0]
        return matches[0]

    item1_s = _pick_non_toc(item1_matches, min_pos=4000)
    item1_start_pos = item1_s.start() if item1_s else 0
    item1a_s = None
    if item1a_matches:
        later_1a = [m for m in item1a_matches if int(m.start()) > int(item1_start_pos + 1000)]
        item1a_s = later_1a[0] if later_1a else _pick_non_toc(item1a_matches, min_pos=6000)
    item1b_s = None
    if item1b_matches:
        later_1b = [m for m in item1b_matches if item1a_s is None or int(m.start()) > int(item1a_s.start())]
        item1b_s = later_1b[0] if later_1b else _pick_non_toc(item1b_matches, min_pos=8000)
    item2_s = _pick_non_toc(item2_matches, min_pos=8000)
    if item1_s:
        s = item1_s.start()
        e = item1a_s.start() if item1a_s and item1a_s.start() > s else min(len(plain), s + 12000)
        biz_text = plain[s:e].strip()
        biz_part = "10-K Item 1"
    else:
        m_biz = re.search(r"\bbusiness\b", low, re.I)
        if m_biz:
            s = max(0, m_biz.start() - 500)
            e = min(len(plain), s + 10000)
            biz_text = plain[s:e].strip()
            biz_part = "10-K Business section (fallback)"
        else:
            biz_text = plain[:9000]
            biz_part = "10-K text fallback"

    risk_text = ""
    risk_part = "10-K Item 1A"
    if item1a_s:
        rs = item1a_s.start()
        re_stop_candidates = []
        if item1b_s and item1b_s.start() > rs:
            re_stop_candidates.append(item1b_s.start())
        if item2_s and item2_s.start() > rs:
            re_stop_candidates.append(item2_s.start())
        re_stop = min(re_stop_candidates) if re_stop_candidates else min(len(plain), rs + 26000)
        risk_text = plain[rs:re_stop].strip()
    else:
        m_risk = re.search(r"\brisk factors?\b", low, re.I)
        if m_risk:
            rs = max(0, m_risk.start() - 200)
            risk_text = plain[rs : min(len(plain), rs + 16000)].strip()
            risk_part = "10-K Risk Factors (fallback)"

    def _split_sentences_local(text: str) -> List[str]:
        txt = re.sub(r"\s+", " ", str(text or "")).strip()
        if not txt:
            return []
        parts = re.split(r"(?<=[\.\!\?])\s+(?=[A-Z0-9])", txt)
        out_sents: List[str] = []
        for p in parts:
            s2 = re.sub(r"\s+", " ", str(p or "")).strip()
            if not s2:
                continue
            if len(s2) < 45:
                continue
            if len(s2) > 560:
                continue
            out_sents.append(s2)
        return out_sents

    company_name = str((submissions or {}).get("name") or (ticker or "The company")).strip()
    profile = get_company_profile(ticker)
    is_pbi_profile = str(getattr(profile, "ticker", "") or ticker or "").strip().upper() == "PBI"
    pbi_summary_description_fallback = str(getattr(profile, "summary_description_fallback", "") or "").strip()
    pbi_summary_key_advantage_fallback = str(
        getattr(profile, "summary_key_advantage_fallback", "") or ""
    ).strip()
    pbi_summary_segment_operating_model_fallbacks = list(
        getattr(profile, "summary_segment_operating_model_fallbacks", tuple()) or tuple()
    )
    pbi_summary_dependency_fallbacks = list(
        getattr(profile, "summary_dependency_fallbacks", tuple()) or tuple()
    )
    pbi_summary_wrong_thesis_fallbacks = list(
        getattr(profile, "summary_wrong_thesis_fallbacks", tuple()) or tuple()
    )
    biz_sentences = _split_sentences_local(biz_text)
    risk_sentences = _split_sentences_local(risk_text)

    boilerplate_re = re.compile(
        r"\b(form\s+10-k\s+summary|forward-looking statements|table of contents|part i\b|item\s*1a|"
        r"risk factors|consolidated financial statements|safe harbor|private securities litigation reform act)\b",
        re.I,
    )

    def _contains_any(text_in: str, terms: Tuple[str, ...]) -> bool:
        low_txt = str(text_in or "").lower()
        return any(str(t).lower() in low_txt for t in terms)

    def _looks_glossary_sentence(text_in: str) -> bool:
        txt_local = re.sub(r"\s+", " ", str(text_in or "")).strip()
        if not txt_local:
            return False
        if re.match(r"^[A-Z]{2,8}\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,6}\b", txt_local):
            return True
        acronym_hits = re.findall(r"\b[A-Z]{2,8}\b", txt_local)
        if len(acronym_hits) >= 3 and not re.search(
            r"\b(provide|provides|produce|produces|operate|operates|sell|sells|market|markets|manufacture|manufactures|process|processes|refine|refines)\b",
            txt_local,
            re.I,
        ):
            return True
        if re.search(r"\bmeans\b|\brefers to\b|\bdefined as\b", txt_local, re.I):
            return True
        return False

    pbi_business_re = re.compile(
        r"\b(sendtech|presort|mail(?:ing|ing technology)?|shipping|parcel|postage|"
        r"customer communications?|shipping software|locker|sorting)\b",
        re.I,
    )
    pbi_business_action_re = re.compile(
        r"\b(provide|provides|offer|offers|enable|enables|include|includes|"
        r"support|supports|serve|serves|deliver|delivers|optimiz|sort|presort)\b",
        re.I,
    )
    pbi_summary_noise_re = re.compile(
        r"\b(table of contents|item\s+\d|part\s+[ivx]+|risk factors?|"
        r"private securities litigation reform act|forward-looking statements|"
        r"see note|page\s+\d+|continued)\b",
        re.I,
    )

    def _pbi_summary_sentence_ok(text_in: str) -> bool:
        txt_local = re.sub(r"\s+", " ", html.unescape(str(text_in or ""))).strip()
        if not txt_local:
            return False
        if pbi_summary_noise_re.search(txt_local):
            return False
        if _looks_glossary_sentence(txt_local):
            return False
        if not pbi_business_re.search(txt_local):
            return False
        return bool(pbi_business_action_re.search(txt_local))

    def _looks_pbi_summary_noise(text_in: str) -> bool:
        txt_local = re.sub(r"\s+", " ", html.unescape(str(text_in or ""))).strip()
        if not txt_local:
            return True
        if pbi_summary_noise_re.search(txt_local):
            return True
        if re.search(
            r"\b(item\s+[0-9a-z]|management'?s discussion|table of contents|"
            r"market for the company'?s common equity|related stockholder matters|"
            r"issuer purchases of equity securities)\b",
            txt_local,
            re.I,
        ):
            return True
        return False

    what_sentences: List[str] = []
    for st in biz_sentences:
        st_clean = re.sub(r"\s+", " ", html.unescape(str(st or ""))).strip()
        st_clean = re.sub(r"^(business\s+overview|overview)\s+", "", st_clean, flags=re.I)
        if not st_clean or boilerplate_re.search(st_clean) or _looks_glossary_sentence(st_clean):
            continue
        if is_pbi_profile and not _pbi_summary_sentence_ok(st_clean):
            continue
        if profile.industry_keywords and not _contains_any(st_clean, profile.industry_keywords):
            continue
        what_sentences.append(st_clean)
        if len(what_sentences) >= 2:
            break
    if not what_sentences:
        for st in biz_sentences:
            st_clean = re.sub(r"\s+", " ", html.unescape(str(st or ""))).strip()
            st_clean = re.sub(r"^(business\s+overview|overview)\s+", "", st_clean, flags=re.I)
            if not st_clean or boilerplate_re.search(st_clean) or _looks_glossary_sentence(st_clean):
                continue
            if is_pbi_profile and not pbi_business_re.search(st_clean):
                continue
            what_sentences.append(st_clean)
            if len(what_sentences) >= 2:
                break
    if what_sentences:
        out["what_it_does"] = " ".join(what_sentences)[:560]
    else:
        out["what_it_does"] = f"{company_name} operates in its disclosed industry and end markets."
    if is_pbi_profile and (
        _looks_pbi_summary_noise(out["what_it_does"])
        or not _pbi_summary_sentence_ok(out["what_it_does"])
        or len(re.findall(r"\b(sendtech|presort|shipping|mailing|postal)\b", str(out["what_it_does"]), re.I)) < 2
    ):
        if pbi_summary_description_fallback:
            out["what_it_does"] = pbi_summary_description_fallback
    out["what_it_does_source"] = f"Source: SEC {form} accn={accn} filed={filed or 'n/a'} ({biz_part})"

    # Segment operating model (one deterministic sentence per major segment when available).
    segment_patterns: List[Tuple[str, re.Pattern[str]]] = list(profile.segment_patterns)
    op_verb_re = re.compile(r"\b(provide|provides|offer|offers|include|includes|generate|serves?|deliver|process(?:es)?)\b", re.I)
    segment_rows: List[Dict[str, Any]] = []
    for seg_name, seg_re in segment_patterns:
        seg_pick = None
        for st in biz_sentences:
            sl = st.lower()
            if not seg_re.search(sl):
                continue
            if not op_verb_re.search(sl):
                continue
            seg_pick = st
            break
        if seg_pick:
            segment_rows.append({"segment": seg_name, "text": seg_pick[:360]})
    if segment_rows:
        out["segment_operating_model"] = segment_rows
        out["segment_operating_model_source"] = (
            f"Source: SEC {form} accn={accn} filed={filed or 'n/a'} ({biz_part})"
        )

    # Concrete key advantage sentence selection from Item 1 when available.
    require_terms = tuple(str(x).lower() for x in profile.key_adv_require_keywords)
    deny_terms = tuple(str(x).lower() for x in profile.key_adv_deny_keywords)
    if biz_sentences:
        scored_adv: List[Tuple[int, str]] = []
        for st in biz_sentences:
            st_clean = re.sub(r"\s+", " ", html.unescape(str(st or ""))).strip()
            sl = st_clean.lower()
            if boilerplate_re.search(sl):
                continue
            if any(deny in sl for deny in deny_terms):
                continue
            req_hits = sum(1 for req in require_terms if req in sl)
            if req_hits < 1:
                continue
            sc = 0
            sc += req_hits * 2
            sc += 2 if re.search(r"\b(advantage|competitive|differentiat|leading|leadership|cost position|scale)\b", sl) else 0
            sc += 1 if re.search(r"\b(customer|client|retention|switching)\b", sl) else 0
            sc += min(2, len(st_clean) // 120)
            scored_adv.append((sc, st_clean))
        if scored_adv:
            scored_adv.sort(key=lambda x: (x[0], len(x[1])), reverse=True)
            out["key_advantage"] = scored_adv[0][1][:360]
            out["key_advantage_source"] = (
                f"Source: SEC {form} accn={accn} filed={filed or 'n/a'} ({biz_part})"
            )
        else:
            out["key_advantage"] = "N/A (no high-confidence competitive advantage sentence found)."
            out["key_advantage_source"] = (
                f"Source: SEC {form} accn={accn} filed={filed or 'n/a'} ({biz_part})"
            )

    # Key dependencies / wrong-thesis bullets from Item 1A risk factors.
    if risk_sentences:
        dep_patterns: List[Tuple[str, re.Pattern[str]]] = []
        if str(profile.ticker or "").upper() == "PBI":
            dep_patterns.extend(
                [
                    ("USPS / postal dependency", re.compile(r"\b(usps|postal service|postal rates?|mailing|postal)\b", re.I)),
                    ("Secular mail-volume decline", re.compile(r"\b(mail volume|declin\w+\s+mail|secular decline)\b", re.I)),
                    ("PB Bank liquidity / trapped capital", re.compile(r"\b(pb bank|trapped capital|bank liquidity|bank funding|funding release)\b", re.I)),
                    ("Cost rationalization execution", re.compile(r"\b(cost savings?|cost rationalization|restructuring|takeout)\b", re.I)),
                    ("SendTech / Presort demand", re.compile(r"\b(sendtech|presort|shipping|parcel|postage)\b", re.I)),
                ]
            )
        dep_patterns.extend(
            [
                ("Commodity / feedstock spread", re.compile(r"\b(corn|natural gas|commodity|spread|margin spread)\b", re.I)),
                ("Policy / regulatory credits", re.compile(r"\b(45z|rin|lcfs|renewable fuel standard|tax credit|regulatory)\b", re.I)),
                ("Carrier / logistics contracts", re.compile(r"\b(carrier|transportation|shipping provider|contract terms?)\b", re.I)),
                ("Leverage / refinancing risk", re.compile(r"\b(debt|leverage|refinanc|maturity|interest rate|covenant)\b", re.I)),
                ("Technology / cybersecurity execution", re.compile(r"\b(cyber|security|technology|systems?|outage|disruption)\b", re.I)),
                ("Macro / demand sensitivity", re.compile(r"\b(macro|economic|inflation|recession|demand)\b", re.I)),
                ("Customer concentration", re.compile(r"\b(customer concentration|significant customer|major customer)\b", re.I)),
            ]
        )
        structural_re = re.compile(r"\b(may|could|might|adversely|materially|depend|subject to)\b", re.I)
        dep_best: Dict[str, Tuple[float, str]] = {}
        for st in risk_sentences:
            st_l = st.lower()
            if is_pbi_profile and pbi_summary_noise_re.search(st_l):
                continue
            for dep_name, dep_re in dep_patterns:
                if not dep_re.search(st_l):
                    continue
                score = 1.0
                score += 1.2 if structural_re.search(st_l) else 0.0
                score += 0.6 if re.search(r"\b(revenue|margin|cash flow|liquidity|debt)\b", st_l) else 0.0
                score += min(0.8, float(len(st)) / 420.0)
                cur = dep_best.get(dep_name)
                if cur is None or score > cur[0]:
                    dep_best[dep_name] = (score, st)
        if dep_best:
            dep_sorted = sorted(dep_best.items(), key=lambda kv: kv[1][0], reverse=True)[:5]
            out["key_dependencies"] = [f"{name}: {txt[:260]}" for name, (_sc, txt) in dep_sorted]
            out["wrong_thesis_bullets"] = [txt[:320] for _name, (_sc, txt) in dep_sorted]
            src_txt = f"Source: SEC {form} accn={accn} filed={filed or 'n/a'} ({risk_part})"
            out["key_dependencies_source"] = src_txt
            out["wrong_thesis_source"] = src_txt
        else:
            fallback_rows: List[str] = []
            for st in risk_sentences:
                st_l = st.lower()
                if not re.search(r"\b(may|could|might|adversely|materially|depend|subject to|risk)\b", st_l):
                    continue
                if not re.search(r"\b(revenue|margin|cash flow|liquidity|debt|customer|postal|regulatory|demand|volume)\b", st_l):
                    continue
                fallback_rows.append(st[:320])
                if len(fallback_rows) >= 5:
                    break
            if not fallback_rows and risk_text:
                risk_low = risk_text.lower()
                dep_hint_patterns = [
                    r"\b(usps|postal|regulatory)\b",
                    r"\b(mail volume|demand|volume)\b",
                    r"\b(debt|leverage|refinanc|maturity|covenant)\b",
                    r"\b(customer concentration|major customer|significant customer)\b",
                    r"\b(cyber|technology|systems?|security)\b",
                ]
                for pat in dep_hint_patterns:
                    m = re.search(pat, risk_low, re.I)
                    if not m:
                        continue
                    s0 = max(0, m.start() - 180)
                    e0 = min(len(risk_text), m.end() + 220)
                    sn = re.sub(r"\s+", " ", risk_text[s0:e0]).strip(" ;:-")
                    if len(sn) < 50:
                        continue
                    fallback_rows.append(sn[:320])
                    if len(fallback_rows) >= 5:
                        break
            if fallback_rows:
                if is_pbi_profile:
                    filtered_fallback_rows = []
                    for row_txt in fallback_rows:
                        row_low = str(row_txt or "").lower()
                        if pbi_summary_noise_re.search(row_low):
                            continue
                        if not re.search(
                            r"\b(usps|postal|mail volume|sendtech|presort|shipping|parcel|"
                            r"debt|leverage|liquidity|pb bank|trapped capital|cost savings?|"
                            r"restructuring|margin|cash flow)\b",
                            row_low,
                            re.I,
                        ):
                            continue
                        filtered_fallback_rows.append(row_txt)
                    fallback_rows = filtered_fallback_rows
            if fallback_rows:
                out["key_dependencies"] = [f"Risk dependency: {x[:260]}" for x in fallback_rows[:5]]
                out["wrong_thesis_bullets"] = fallback_rows[:5]
                src_txt = f"Source: SEC {form} accn={accn} filed={filed or 'n/a'} ({risk_part})"
                out["key_dependencies_source"] = src_txt
                out["wrong_thesis_source"] = src_txt

    if is_pbi_profile:
        if (
            not out.get("segment_operating_model")
            or len(out.get("segment_operating_model") or []) < 2
            or all(
                "represents approximately" in str((row or {}).get("text") or "").lower()
                for row in list(out.get("segment_operating_model") or [])
            )
        ) and pbi_summary_segment_operating_model_fallbacks:
            seg_rows: List[Dict[str, Any]] = []
            for raw_row in pbi_summary_segment_operating_model_fallbacks[:4]:
                row_txt = re.sub(r"\s+", " ", str(raw_row or "")).strip()
                if not row_txt:
                    continue
                seg_name = row_txt.split(":", 1)[0].strip()
                seg_rows.append({"segment": seg_name, "text": row_txt[:360]})
            if seg_rows:
                out["segment_operating_model"] = seg_rows
                out["segment_operating_model_source"] = "Source: Profile fallback (clean annual segment description unavailable)"

        if (
            not str(out.get("key_advantage") or "").strip()
            or str(out.get("key_advantage") or "").strip().startswith("N/A")
            or _looks_pbi_summary_noise(str(out.get("key_advantage") or ""))
            or not re.search(
                r"\b(installed base|network|software|workflow|recurring|presort|shipping|mailing|retention)\b",
                str(out.get("key_advantage") or ""),
                re.I,
            )
        ) and pbi_summary_key_advantage_fallback:
            out["key_advantage"] = pbi_summary_key_advantage_fallback
            out["key_advantage_source"] = "Source: Profile fallback (clean annual competitive-position sentence unavailable)"

        dep_blob = " | ".join(out.get("key_dependencies") or [])
        wrong_blob = " | ".join(out.get("wrong_thesis_bullets") or [])
        pbi_dep_hits = sum(
            1
            for row in list(out.get("key_dependencies") or [])
            if re.search(
                r"\b(usps|postal|mail volume|sendtech|presort|shipping|parcel|pb bank|trapped capital|"
                r"liquidity|cost savings?|restructuring|leverage|debt|margin|cash flow)\b",
                str(row or ""),
                re.I,
            )
        )
        pbi_wrong_hits = sum(
            1
            for row in list(out.get("wrong_thesis_bullets") or [])
            if re.search(
                r"\b(mail volume|sendtech|presort|pb bank|trapped capital|liquidity|cost savings?|"
                r"service levels?|retention|margin|leverage|refinanc)\b",
                str(row or ""),
                re.I,
            )
        )
        dep_needs_fallback = (
            not out.get("key_dependencies")
            or len(out.get("key_dependencies") or []) < 3
            or pbi_dep_hits < 3
            or bool(re.search(r"\b(cyber threats|u\.s\. government contractor|favorable postage rates are reversed|worldwide customs)\b", dep_blob, re.I))
        )
        wrong_needs_fallback = (
            not out.get("wrong_thesis_bullets")
            or len(out.get("wrong_thesis_bullets") or []) < 3
            or pbi_wrong_hits < 3
            or bool(re.search(r"\b(cyber threats|u\.s\. government contractor|favorable postage rates are reversed|worldwide customs)\b", wrong_blob, re.I))
        )
        if dep_needs_fallback and pbi_summary_dependency_fallbacks:
            out["key_dependencies"] = list(pbi_summary_dependency_fallbacks[:5])
        if wrong_needs_fallback and pbi_summary_wrong_thesis_fallbacks:
            out["wrong_thesis_bullets"] = list(pbi_summary_wrong_thesis_fallbacks[:5])

    # Revenue streams table extraction from latest 10-K tables
    report_year = report.year if isinstance(report, date) else None
    best_streams: List[Dict[str, Any]] = []
    best_score = -1
    best_table_idx = None
    tables = read_html_tables_any(html_bytes)
    for tidx, tdf in enumerate(tables):
        if tdf is None or tdf.empty:
            continue
        df = tdf.copy()
        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
        if df.shape[0] < 3 or df.shape[1] < 2:
            continue
        table_text = re.sub(r"\s+", " ", " ".join([str(x) for x in df.fillna("").astype(str).values.flatten()])).lower()
        first_col = df.columns[0]
        labels = (
            df[first_col]
            .astype(str)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )
        if labels.replace("", pd.NA).dropna().nunique() < 2:
            continue
        label_low = labels.str.lower()
        numeric_cols: List[Any] = []
        for col in df.columns[1:]:
            vals = df[col].apply(lambda x: coerce_number(x))
            valid = vals.notna().sum()
            if valid >= max(2, int(len(df) * 0.35)):
                numeric_cols.append(col)
        if not numeric_cols:
            continue
        col_pick = numeric_cols[-1]
        if report_year is not None:
            for col in numeric_cols:
                if re.search(rf"\b{int(report_year)}\b", str(col)):
                    col_pick = col
                    break
        vals_pick = df[col_pick].apply(lambda x: coerce_number(x))
        total_mask = label_low.str.contains(r"\btotal\b", regex=True)
        total_val = None
        if total_mask.any():
            tv = pd.to_numeric(vals_pick[total_mask], errors="coerce").dropna()
            if not tv.empty:
                total_val = float(tv.iloc[0])
        if total_val in (None, 0):
            pos_sum = pd.to_numeric(vals_pick, errors="coerce").dropna()
            pos_sum = pos_sum[pos_sum > 0]
            total_val = float(pos_sum.sum()) if not pos_sum.empty else None
        if total_val in (None, 0):
            continue
        cand_rows: List[Dict[str, Any]] = []
        non_revenue_label_hits = 0
        for lbl, vv in zip(labels.tolist(), vals_pick.tolist()):
            if vv is None or pd.isna(vv):
                continue
            lbl_s = str(lbl).strip()
            lbl_l = lbl_s.lower()
            if not lbl_s:
                continue
            if re.fullmatch(r"[$€£]", lbl_s):
                continue
            if "revenues from external customers" in lbl_l:
                continue
            if any(x in lbl_l for x in ["total", "elimination", "intersegment", "inter-company", "intercompany"]):
                continue
            if any(
                x in lbl_l
                for x in [
                    "interest",
                    "tax",
                    "provision",
                    "income",
                    "expense",
                    "depreciation",
                    "amortization",
                    "ebit",
                    "earnings",
                    "gain",
                    "loss",
                    "operating",
                    "cost of",
                    "gross profit",
                ]
            ):
                non_revenue_label_hits += 1
                continue
            if any(x in lbl_l for x in ["asset", "liabil", "cash", "investment", "goodwill", "intangible", "receivable", "deferred"]):
                non_revenue_label_hits += 1
                continue
            amt = float(vv)
            if amt <= 0:
                continue
            pct = amt / float(total_val)
            cand_rows.append({"name": lbl_s, "amount": amt, "pct": pct})
        if len(cand_rows) < 2:
            continue
        pct_sum = float(sum(r["pct"] for r in cand_rows))
        hdr_txt = " ".join([str(c) for c in df.columns]).lower()
        revenue_relevant = (
            ("revenue" in hdr_txt)
            or ("disaggregation" in hdr_txt)
            or ("net sales" in hdr_txt)
            or ("revenue" in table_text)
        )
        if not revenue_relevant:
            continue
        score = 0
        if "revenue" in hdr_txt or "disaggregation" in hdr_txt or "segment" in hdr_txt:
            score += 3
        if report_year is not None and re.search(rf"\b{int(report_year)}\b", hdr_txt):
            score += 2
        if total_mask.any():
            score += 2
        if 0.80 <= pct_sum <= 1.20:
            score += 2
        profile_seg_label_hit = False
        for _seg_nm, _seg_re in segment_patterns:
            try:
                if label_low.apply(lambda _v: bool(_seg_re.search(str(_v)))).any():
                    profile_seg_label_hit = True
                    break
            except Exception:
                continue
        if profile_seg_label_hit or label_low.str.contains("segment|service|software|fuel|ethanol|agribusiness|protein", regex=True).any():
            score += 1
        if non_revenue_label_hits > 0:
            score -= 2
        if score > best_score:
            best_score = score
            best_table_idx = tidx
            best_streams = cand_rows

    if best_streams and best_score >= 4:
        best_streams = sorted(best_streams, key=lambda x: float(x.get("pct") or 0), reverse=True)
        if len(best_streams) > 6:
            head = best_streams[:5]
            tail = best_streams[5:]
            head.append(
                {
                    "name": "Other",
                    "amount": float(sum(float(t.get("amount") or 0.0) for t in tail)),
                    "pct": float(sum(float(t.get("pct") or 0.0) for t in tail)),
                }
            )
            best_streams = head
        out["revenue_streams"] = best_streams
        out["revenue_streams_period"] = report
        out["revenue_streams_source"] = (
            f"Source: SEC {form} accn={accn} filed={filed or 'n/a'} "
            f"(10-K revenue/segment table #{int(best_table_idx) + 1})"
        )
    else:
        # Fallback: latest 10-Q segment results table.
        tenq_rows: List[Dict[str, Any]] = []
        for batch in _iter_submission_batches(sec, submissions):
            forms = batch.get("form", []) or []
            accns = batch.get("accessionNumber", []) or []
            report_dates = batch.get("reportDate", []) or []
            filing_dates = batch.get("filingDate", []) or []
            primary_docs = batch.get("primaryDocument", []) or []
            n = min(len(forms), len(accns))
            for i in range(n):
                f = str(forms[i] or "").upper().strip()
                if not f.startswith("10-Q"):
                    continue
                a = str(accns[i] or "").strip()
                d = str(primary_docs[i] or "") if i < len(primary_docs) else ""
                if not a or not d:
                    continue
                tenq_rows.append(
                    {
                        "form": f,
                        "accn": a,
                        "doc": d,
                        "filed": parse_date(filing_dates[i]) if i < len(filing_dates) else None,
                        "report": parse_date(report_dates[i]) if i < len(report_dates) else None,
                    }
                )

        def _extract_segment_streams_from_tables(
            src_tables: List[pd.DataFrame],
            target_year: Optional[int],
        ) -> Tuple[List[Dict[str, Any]], Optional[int]]:
            best_rows: List[Dict[str, Any]] = []
            best_idx: Optional[int] = None
            best_sc = -1
            for tidx2, tdf2 in enumerate(src_tables):
                if tdf2 is None or tdf2.empty:
                    continue
                dfx = tdf2.copy().dropna(how="all").dropna(axis=1, how="all")
                if dfx.shape[0] < 4 or dfx.shape[1] < 2:
                    continue
                table_txt = re.sub(r"\s+", " ", " ".join(str(x) for x in dfx.fillna("").astype(str).values.flatten())).lower()
                profile_segment_hit = any(seg_re.search(table_txt) for _seg_nm, seg_re in segment_patterns)
                if not (
                    "segment revenue" in table_txt
                    or "segment results" in table_txt
                    or profile_segment_hit
                ):
                    continue

                cols = list(dfx.columns)
                data_cols = cols[1:]
                if not data_cols:
                    continue
                first_labels = [re.sub(r"\s+", " ", str(v)).strip().lower() for v in dfx[cols[0]].tolist()]
                has_total_segment_row = any("total segment revenue" in ll for ll in first_labels)
                year_by_col: Dict[Any, int] = {}
                for ridx2 in range(min(5, len(dfx))):
                    for c in data_cols:
                        mv = re.search(r"\b(20\d{2})\b", str(dfx.iloc[ridx2][c]))
                        if mv:
                            year_by_col[c] = int(mv.group(1))
                numeric_cols = []
                for c in data_cols:
                    vals = dfx[c].apply(coerce_number)
                    if vals.notna().sum() >= 2:
                        numeric_cols.append(c)
                if not numeric_cols:
                    continue
                col_pick = numeric_cols[0]
                if target_year is not None:
                    cands = [c for c in numeric_cols if year_by_col.get(c) == int(target_year)]
                    if cands:
                        col_pick = cands[0]

                seg_map: Dict[str, float] = {}
                total_val: Optional[float] = None
                for ridx2 in range(len(dfx)):
                    label = str(dfx.iloc[ridx2][cols[0]]).strip()
                    if not label:
                        continue
                    ll = re.sub(r"\s+", " ", label).lower()
                    if re.fullmatch(r"[$€£]", label):
                        continue
                    if "revenues from external customers" in ll:
                        continue
                    v = coerce_number(dfx.iloc[ridx2][col_pick])
                    if v is None:
                        continue
                    vv = float(v)
                    if vv <= 0:
                        continue
                    if "total segment revenue" in ll or ll == "total revenue" or ll.endswith(" total revenue"):
                        total_val = vv
                        continue
                    if any(k in ll for k in ["other", "total", "elimination", "intersegment", "subtotal", "timing of revenue"]):
                        continue
                    if any(
                        k in ll
                        for k in [
                            "revenue",
                            "less",
                            "cost of",
                            "operating expenses",
                            "adjusted",
                            "interest",
                            "corporate",
                            "reconciliation",
                        ]
                    ):
                        continue
                    if any(seg_re.search(ll) for _seg_nm, seg_re in segment_patterns):
                        seg_map[label] = max(seg_map.get(label, 0.0), vv)
                    elif any(k in ll for k in ["segment", "services", "solutions", "fuel", "ethanol", "agribusiness", "protein"]):
                        seg_map[label] = max(seg_map.get(label, 0.0), vv)
                    elif has_total_segment_row and re.match(r"^[A-Za-z][A-Za-z&/ \-]{2,40}$", label):
                        seg_map[label] = max(seg_map.get(label, 0.0), vv)

                if len(seg_map) < 2:
                    continue
                if total_val in (None, 0):
                    total_val = float(sum(seg_map.values()))
                if total_val in (None, 0):
                    continue
                rows2 = [{"name": k, "amount": v, "pct": (v / float(total_val))} for k, v in seg_map.items()]
                rows2 = sorted(rows2, key=lambda x: float(x.get("pct") or 0.0), reverse=True)
                pct_sum = float(sum(float(r.get("pct") or 0.0) for r in rows2))
                sc = 0
                if "segment revenue" in table_txt:
                    sc += 4
                if has_total_segment_row:
                    sc += 6
                if "sendtech" in table_txt and "presort" in table_txt:
                    sc += 3
                if target_year is not None and year_by_col and any(y == target_year for y in year_by_col.values()):
                    sc += 2
                if 0.80 <= pct_sum <= 1.20:
                    sc += 2
                if total_val and total_val > 100_000:
                    sc += 1
                if sc > best_sc:
                    best_sc = sc
                    best_rows = rows2
                    best_idx = tidx2
            return best_rows, best_idx

        def _extract_segment_streams_from_inline(
            html_text_q: str,
            q_end: Optional[date],
        ) -> List[Dict[str, Any]]:
            if not html_text_q or q_end is None or BeautifulSoup is None:
                return []
            try:
                soup_q = BeautifulSoup(html_text_q, "html.parser")
            except Exception:
                return []

            if q_end.month not in (3, 6, 9, 12):
                return []
            if q_end.month == 3:
                q_start = date(q_end.year, 1, 1)
            else:
                q_start = date(q_end.year, q_end.month - 2, 1)
            q_end_s = q_end.strftime("%Y-%m-%d")
            q_start_s = q_start.strftime("%Y-%m-%d")

            contexts: Dict[str, Dict[str, Any]] = {}
            for ctx in soup_q.find_all(re.compile(r".*context$", re.I)):
                cid = str(ctx.get("id") or "").strip()
                if not cid:
                    continue
                members = ctx.find_all(re.compile(r".*explicitmember$", re.I))
                dims: List[Tuple[str, str]] = []
                seg_member = ""
                for m in members:
                    dim = str(m.get("dimension") or "").strip()
                    val = m.get_text(" ", strip=True)
                    dims.append((dim, val))
                    if dim.endswith("StatementBusinessSegmentsAxis"):
                        seg_member = val
                if not seg_member:
                    continue
                sd = ctx.find(re.compile(r".*startdate$", re.I))
                ed = ctx.find(re.compile(r".*enddate$", re.I))
                if not (sd and ed):
                    continue
                start_s = sd.get_text(strip=True)
                end_s = ed.get_text(strip=True)
                if start_s != q_start_s or end_s != q_end_s:
                    continue
                contexts[cid] = {"segment": seg_member, "dims": dims}

            if not contexts:
                return []

            def _clean_seg(member: str) -> str:
                nm = str(member or "").split(":")[-1]
                nm = re.sub(r"Member$", "", nm)
                nm = re.sub(r"([a-z])([A-Z])", r"\1 \2", nm)
                nm = nm.replace("  ", " ").strip()
                return nm or str(member)

            seen_fact_keys: set[Tuple[str, str, float]] = set()
            seg_vals: Dict[str, float] = {}
            for fact in soup_q.find_all(re.compile(r".*nonfraction$", re.I)):
                ctx = str(fact.get("contextref") or fact.get("contextRef") or "").strip()
                if ctx not in contexts:
                    continue
                name = str(fact.get("name") or "")
                name_l = name.lower()
                if not (
                    name_l.endswith(":revenues")
                    or "revenuefromcontractwithcustomerexcludingassessedtax" in name_l
                ):
                    continue
                dims = contexts[ctx]["dims"]
                if len(dims) > 2:
                    continue
                txt = fact.get_text(" ", strip=True).replace("\xa0", " ")
                neg = "(" in txt and ")" in txt
                num_txt = re.sub(r"[^0-9\.\-]", "", txt)
                if not num_txt:
                    continue
                try:
                    val = float(num_txt)
                except Exception:
                    continue
                if neg:
                    val = -abs(val)
                if val <= 0:
                    continue
                scale_raw = fact.get("scale")
                try:
                    scale_i = int(scale_raw) if scale_raw is not None else 0
                except Exception:
                    scale_i = 0
                val = float(val) * (10 ** scale_i)
                seg_name = _clean_seg(contexts[ctx]["segment"])
                k = (ctx, seg_name, round(val, 4))
                if k in seen_fact_keys:
                    continue
                seen_fact_keys.add(k)
                # Keep max per segment to avoid double-counting repeated presentation lines.
                seg_vals[seg_name] = max(seg_vals.get(seg_name, 0.0), val)

            if len(seg_vals) < 2:
                return []
            total = float(sum(v for v in seg_vals.values() if v > 0))
            if total <= 0:
                return []
            rows = [
                {"name": k, "amount": float(v), "pct": float(v) / total}
                for k, v in seg_vals.items()
                if v > 0
            ]
            rows = sorted(rows, key=lambda x: float(x.get("pct") or 0.0), reverse=True)
            return rows

        if tenq_rows:
            tenq_rows = sorted(
                tenq_rows,
                key=lambda r: (
                    pd.Timestamp(r.get("filed")) if r.get("filed") is not None else pd.Timestamp("1900-01-01"),
                    pd.Timestamp(r.get("report")) if r.get("report") is not None else pd.Timestamp("1900-01-01"),
                ),
                reverse=True,
            )
            qsel = tenq_rows[0]
            q_accn = str(qsel.get("accn") or "")
            q_doc = str(qsel.get("doc") or "")
            q_form = str(qsel.get("form") or "10-Q")
            q_filed = qsel.get("filed")
            q_report = qsel.get("report")
            try:
                q_html = sec.download_document(cik_int, normalize_accession(q_accn), q_doc)
                q_html_text = q_html.decode("utf-8", errors="ignore")
                q_rows = _extract_segment_streams_from_inline(q_html_text, q_report if isinstance(q_report, date) else None)
                q_tidx = None
                if not q_rows:
                    q_tables = read_html_tables_any(q_html)
                    q_rows, q_tidx = _extract_segment_streams_from_tables(q_tables, q_report.year if isinstance(q_report, date) else None)
            except Exception:
                q_rows, q_tidx = ([], None)
            if q_rows:
                out["revenue_streams"] = q_rows
                out["revenue_streams_period"] = q_report
                if q_tidx is None:
                    out["revenue_streams_source"] = (
                        f"Source: SEC {q_form} accn={q_accn} filed={q_filed or 'n/a'} "
                        "(inline segment revenue facts)"
                    )
                else:
                    out["revenue_streams_source"] = (
                        f"Source: SEC {q_form} accn={q_accn} filed={q_filed or 'n/a'} "
                        f"(Segment Results revenue table #{int(q_tidx) + 1})"
                    )
            else:
                out["revenue_streams"] = []
                out["revenue_streams_source"] = (
                    f"Source: SEC {form} accn={accn} filed={filed or 'n/a'} "
                    "(N/A: not found in parsed 10-K/10-Q segment tables)"
                )
        else:
            out["revenue_streams"] = []
            out["revenue_streams_source"] = (
                f"Source: SEC {form} accn={accn} filed={filed or 'n/a'} "
                "(N/A: not found in parsed 10-K tables)"
            )

    if not out.get("segment_operating_model") and isinstance(out.get("revenue_streams"), list):
        seg_fallback: List[Dict[str, Any]] = []
        for rr in list(out.get("revenue_streams") or [])[:3]:
            nm = str(rr.get("name") or "").strip()
            if not nm:
                continue
            if profile.segment_patterns:
                nm_l = nm.lower()
                if not any(seg_re.search(nm_l) for _seg_name, seg_re in profile.segment_patterns):
                    continue
            pct = pd.to_numeric(rr.get("pct"), errors="coerce")
            if pd.notna(pct):
                txt = f"{nm} represents approximately {float(pct) * 100.0:.1f}% of reported revenue mix."
            else:
                txt = f"{nm} is a material operating revenue stream."
            seg_fallback.append({"segment": nm, "text": txt})
        if seg_fallback:
            out["segment_operating_model"] = seg_fallback
            out["segment_operating_model_source"] = str(out.get("revenue_streams_source") or "")

    return out


def build_revolver_availability(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_docs: int = 12,
    lookback_years: int = 5,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    def _safe_compile(pat: str, flags: int = re.I, fallback: Optional[str] = None) -> re.Pattern:
        try:
            return re.compile(pat, flags)
        except re.error:
            if fallback:
                return re.compile(fallback, flags)
            # very permissive fallback to avoid hard crash
            return re.compile(r"revolver", re.I)

    revolver_re = _safe_compile(r"revolver|revolving credit facility|revolving facility", re.I)
    amount_re = _safe_compile(
        r"(?:\$\s*)?([0-9]{1,3}(?:,[0-9]{3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)(?:\s*(million|billion))?",
        re.I,
    )
    agg_re = _safe_compile(
        r"aggregate\s+commitments[^0-9]{0,60}(?:\$\s*)?([0-9]{1,3}(?:,[0-9]{3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)(?:\s*(million|billion))?",
        re.I,
    )
    rev_re1 = _safe_compile(
        r"(?:\$\s*)?([0-9]{1,3}(?:,[0-9]{3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)(?:\s*(million|billion))?\s+revolving\s+credit\s+facility",
        re.I,
    )
    rev_re2 = _safe_compile(
        r"revolving\s+credit\s+facility[^0-9]{0,40}(?:\$\s*)?([0-9]{1,3}(?:,[0-9]{3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)(?:\s*(million|billion))?",
        re.I,
    )
    perm_re = _safe_compile(
        r"permitted\s+borrowings[^0-9]{0,60}(?:\$\s*)?([0-9]{1,3}(?:,[0-9]{3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)(?:\s*(million|billion))?",
        re.I,
    )
    capacity_re = _safe_compile(
        r"(?:borrowing\s+capacity|capacity\s+under\s+the\s+revolving\s+credit\s+facility)[^0-9]{0,60}"
        r"(?:\$\s*)?([0-9]{1,3}(?:,[0-9]{3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)(?:\s*(million|billion))?",
        re.I,
    )
    capacity_re = _safe_compile(
        r"(?:borrowing\s+capacity|capacity\s+under\s+the\s+revolving\s+credit\s+facility)[^0-9]{0,60}"
        r"(?:\$\s*)?([0-9]{1,3}(?:,[0-9]{3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)(?:\s*(million|billion))?",
        re.I,
    )
    provides_re = _safe_compile(
        r"provides?\s+for[^0-9]{0,60}(?:\$\s*)?([0-9]{1,3}(?:,[0-9]{3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)(?:\s*(million|billion))?",
        re.I,
    )
    keyword_l = tuple(k.lower() for k in REVOLVER_SCAN_KEYWORDS)

    def _contains_revolver_keyword(text: str) -> bool:
        if not text:
            return False
        t = text.lower()
        return any(k in t for k in keyword_l)

    def _doc_filter_reason(doc_name: str, *, is_primary: bool = False) -> Tuple[bool, str]:
        nm = (doc_name or "").strip()
        if not nm:
            return False, "empty"
        low = nm.lower()
        if re.search(r"\.(xsd|xml|json|zip|jpg|jpeg|png|gif|xls|xlsx)$", low, re.I):
            return False, "non-html"
        file_has_kw = _contains_revolver_keyword(low) or bool(REVOLVER_DOC_ALLOW_RE.search(low))
        denied = bool(REVOLVER_DOC_DENY_RE.search(low))
        if is_primary:
            if denied and not file_has_kw:
                return False, "denylist-primary-no-keyword"
            return True, "primary"
        if denied and not file_has_kw:
            return False, "denylist/no-keyword"
        if file_has_kw:
            return True, "allowlist/keyword"
        return False, "no-keyword"

    def _quick_scan_text(html_text: str) -> str:
        if not html_text:
            return ""
        cleaned = re.sub(r"<script[^>]*>.*?</script>", " ", html_text, flags=re.I | re.S)
        cleaned = re.sub(r"<style[^>]*>.*?</style>", " ", cleaned, flags=re.I | re.S)
        cleaned = re.sub(r"<!--.*?-->", " ", cleaned, flags=re.S)
        cleaned = re.sub(r"<[^>]+>", " ", cleaned)
        cleaned = html.unescape(cleaned).replace("\xa0", " ")
        return re.sub(r"\s+", " ", cleaned).strip()

    def _table_is_revolver_candidate(table_text: str) -> bool:
        if not table_text:
            return False
        low = table_text.lower()
        if _contains_revolver_keyword(low):
            return True
        hint_hits = sum(1 for h in REVOLVER_TABLE_HINTS if h in low)
        has_num = bool(re.search(r"\b\d[\d,\.]*\b", low))
        return hint_hits >= 2 and has_num

    def _extract_table_blocks(html_text: str) -> List[str]:
        if not html_text:
            return []
        if BeautifulSoup is not None:
            try:
                soup = BeautifulSoup(html_text, "html.parser")
                return [str(t) for t in soup.find_all("table")]
            except Exception:
                pass
        return re.findall(r"<table\b[^>]*>.*?</table>", html_text, flags=re.I | re.S)

    def _load_revolver_doc_cache() -> Tuple[Path, Dict[str, Any]]:
        cache_path = sec.cache_dir / "revolver_doc_cache.json"
        try:
            if cache_path.exists():
                payload = json.loads(cache_path.read_text(encoding="utf-8"))
                if isinstance(payload, dict) and payload.get("version") == REVOLVER_CACHE_VERSION:
                    docs = payload.get("docs")
                    if isinstance(docs, dict):
                        return cache_path, docs
        except Exception:
            pass
        return cache_path, {}

    def _dump_revolver_doc_cache(cache_path: Path, docs_cache: Dict[str, Any]) -> None:
        try:
            payload = {"version": REVOLVER_CACHE_VERSION, "docs": docs_cache}
            cache_path.write_text(json.dumps(payload, ensure_ascii=True), encoding="utf-8")
        except Exception:
            pass

    def _json_safe(v: Any) -> Any:
        if isinstance(v, (dt.date, datetime)):
            return v.isoformat()
        if isinstance(v, pd.Timestamp):
            return v.date().isoformat()
        if hasattr(v, "item"):
            try:
                return v.item()
            except Exception:
                pass
        return v

    def _snippet(txt: str, start: int, end: int, window: int = 260) -> str:
        s = max(0, start - window)
        e = min(len(txt), end + window)
        return re.sub(r"\s+", " ", txt[s:e]).strip()

    def _near_revolver(txt: str, pos: int, window: int = 120) -> bool:
        if not txt:
            return False
        s = max(0, pos - window)
        e = min(len(txt), pos + window)
        return bool(revolver_re.search(txt[s:e]))

    def _amount_ok(ctx: str, match_text: str, pos: Optional[int] = None) -> bool:
        if not ctx or not match_text:
            return False
        if pos is None:
            idx = ctx.lower().find(match_text.lower())
        else:
            idx = pos
        if idx is None or idx < 0:
            idx = 0
        if not _near_revolver(ctx, idx, window=120):
            return False
        bad = re.search(
            r"term\s+loan|swap|interest\s+rate\s+swap|convertible|notes?\s+due|senior\s+notes|debenture|bond",
            ctx[max(0, idx - 90): idx + 90],
            re.I,
        )
        return bad is None

    def _scale_num(num_str: str, unit: str, ctx: str, match_text: Optional[str] = None) -> Optional[float]:
        raw = num_str
        num = float(raw.replace(",", ""))
        if 1900 <= num <= 2100 and "," not in raw:
            return None
        if match_text and "&#" in match_text and num < 10000:
            # skip HTML entity artifacts (e.g., &#8217;)
            return None
        unit_l = (unit or "").lower()
        match_ctx = match_text or ctx
        has_scale_word = bool(re.search(r"\b(million|billion)\b", match_ctx, re.I))
        has_dollar = "$" in match_ctx
        if unit_l.startswith("b"):
            num *= 1e9
        elif unit_l.startswith("m"):
            num *= 1e6
        else:
            if has_scale_word:
                if re.search(r"\bbillion\b", ctx, re.I):
                    num *= 1e9
                else:
                    num *= 1e6 if num < 1e6 else 1.0
            else:
                if num < 1e6:
                    return None
        if num < 1e6:
            return None
        if num < 10_000_000 and not has_dollar and not has_scale_word:
            return None
        return num

    def _find_commitment_with_method(ctx: str) -> Tuple[Optional[float], Optional[str], Optional[float], Optional[str]]:
        # Returns (commitment, commit_method, facility_size, facility_method)
        facility: Optional[float] = None
        facility_method: Optional[str] = None
        # Capture facility size first; commitment logic can then decide whether to reuse it.
        for pat in (rev_re1, rev_re2):
            m = pat.search(ctx)
            if not m:
                continue
            v = _scale_num(m.group(1), m.group(2) or "", ctx, m.group(0))
            if v is not None:
                facility = v
                facility_method = "facility_size"
                break
        chg_generic = re.search(
            r"(?:permitted\s+borrowings|revolver\s+commitments?|aggregate\s+(?:amount\s+of\s+)?(?:revolver\s+)?commitments?|revolving\s+credit\s+facility)"
            r"[^\.]{0,220}?from\s+\$?\s*([0-9][0-9,\.]*)\s*(million|billion)?"
            r"[^\.]{0,80}?to\s+\$?\s*([0-9][0-9,\.]*)\s*(million|billion)?",
            ctx,
            re.I,
        )
        if chg_generic and revolver_re.search(ctx):
            unit = chg_generic.group(4) or chg_generic.group(2) or ""
            v2 = _scale_num(chg_generic.group(3), unit, ctx, chg_generic.group(0))
            if v2 is not None:
                return v2, "change_phrase", facility, facility_method
        inc1 = re.search(
            r"increase(?:d)?\s+from\s+\$?\s*([0-9][0-9,\.]*)\s*(million|billion)?\s+to\s+\$?\s*([0-9][0-9,\.]*)\s*(million|billion)?",
            ctx,
            re.I,
        )
        if inc1 and revolver_re.search(ctx) and _amount_ok(ctx, inc1.group(0), inc1.start()):
            unit = inc1.group(4) or inc1.group(2) or ""
            v2 = _scale_num(inc1.group(3), unit, ctx, inc1.group(0))
            if v2 is not None:
                return v2, "change_phrase", facility, facility_method
        inc2 = re.search(
            r"increase(?:d)?\s+to\s+\$?\s*([0-9][0-9,\.]*)\s*(million|billion)?\s+from\s+\$?\s*([0-9][0-9,\.]*)\s*(million|billion)?",
            ctx,
            re.I,
        )
        if inc2 and revolver_re.search(ctx) and _amount_ok(ctx, inc2.group(0), inc2.start()):
            unit = inc2.group(2) or inc2.group(4) or ""
            v2 = _scale_num(inc2.group(1), unit, ctx, inc2.group(0))
            if v2 is not None:
                return v2, "change_phrase", facility, facility_method
        dec1 = re.search(
            r"(?:reduce|reduces|reduced|decrease|decreases|decreased|lower(?:ed)?)\s+from\s+\$?\s*([0-9][0-9,\.]*)\s*(million|billion)?\s+to\s+\$?\s*([0-9][0-9,\.]*)\s*(million|billion)?",
            ctx,
            re.I,
        )
        if dec1 and revolver_re.search(ctx) and _amount_ok(ctx, dec1.group(0), dec1.start()):
            unit = dec1.group(4) or dec1.group(2) or ""
            v2 = _scale_num(dec1.group(3), unit, ctx, dec1.group(0))
            if v2 is not None:
                return v2, "change_phrase", facility, facility_method
        dec2 = re.search(
            r"(?:reduce|reduces|reduced|decrease|decreases|decreased|lower(?:ed)?)\s+to\s+\$?\s*([0-9][0-9,\.]*)\s*(million|billion)?\s+from\s+\$?\s*([0-9][0-9,\.]*)\s*(million|billion)?",
            ctx,
            re.I,
        )
        if dec2 and revolver_re.search(ctx) and _amount_ok(ctx, dec2.group(0), dec2.start()):
            unit = dec2.group(2) or dec2.group(4) or ""
            v2 = _scale_num(dec2.group(1), unit, ctx, dec2.group(0))
            if v2 is not None:
                return v2, "change_phrase", facility, facility_method
        m_prov = provides_re.search(ctx)
        # "provides for $X revolving credit facility" can appear in the same sentence as term loans.
        if m_prov and revolver_re.search(ctx) and _near_revolver(ctx, m_prov.start(), window=180):
            v2 = _scale_num(m_prov.group(1), m_prov.group(2) or "", ctx, m_prov.group(0))
            if v2 is not None:
                return v2, "provides_for", facility, facility_method
        for pat in (agg_re,):
            m = pat.search(ctx)
            if not m:
                continue
            if pat is agg_re and not revolver_re.search(ctx):
                continue
            if not _amount_ok(ctx, m.group(0), m.start()):
                continue
            v = _scale_num(m.group(1), m.group(2) or "", ctx, m.group(0))
            if v is not None:
                return v, "aggregate_commitments", facility, facility_method
        m_perm = perm_re.search(ctx)
        if m_perm and revolver_re.search(ctx) and _amount_ok(ctx, m_perm.group(0), m_perm.start()):
            v = _scale_num(m_perm.group(1), m_perm.group(2) or "", ctx, m_perm.group(0))
            if v is not None:
                return v, "permitted_borrowings", facility, facility_method
        m_cap = capacity_re.search(ctx)
        if m_cap and revolver_re.search(ctx) and _amount_ok(ctx, m_cap.group(0), m_cap.start()):
            v = _scale_num(m_cap.group(1), m_cap.group(2) or "", ctx, m_cap.group(0))
            if v is not None:
                return v, "borrowing_capacity", facility, facility_method
        return None, None, facility, facility_method

    def _find_commitment(ctx: str) -> Optional[float]:
        v, _m, _f, _fm = _find_commitment_with_method(ctx)
        return v

    def _is_future_event(snip: str, q_end: Optional[dt.date]) -> bool:
        if q_end is None:
            return False
        # If snippet explicitly anchors to an "as of/at" date <= quarter end, treat as not future
        asof = _asof_date_in_text(snip)
        if asof and asof <= q_end:
            return False
        # Look for "In February 2025" style references
        m = re.search(r"\bin\s+(january|february|march|april|may|june|july|august|september|october|november|december)\s+(20\d{2})", snip, re.I)
        if not m:
            return False
        month = m.group(1).lower()
        year = int(m.group(2))
        month_num = {
            "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
            "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
        }.get(month, 0)
        if year > q_end.year:
            return True
        if year == q_end.year and month_num > q_end.month:
            return True
        return False

    def _asof_date_in_text(snip: str) -> Optional[dt.date]:
        m = re.search(
            r"\b(?:as\s+of|at)\s+((?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|"
            r"Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\.?\s+\d{1,2},?\s+\d{4})",
            snip,
            re.I,
        )
        if not m:
            return None
        try:
            return parse_date(m.group(1))
        except Exception:
            return None

    def _score_snip(snip: str, commitment: Optional[float], drawn: Optional[float], availability: Optional[float]) -> int:
        s = snip.lower()
        score = 0
        if "revolv" in s or "revolving credit facility" in s:
            score += 3
        if "new credit agreement" in s or "entered into a new" in s:
            score += 6
        if "aggregate commitments" in s or "total commitments" in s or "commitment" in s:
            score += 3
        if re.search(r"increase|increased|amend|amended|upsiz|expand|reduce|reduced|decrease|decreased|lowered", s):
            score += 4
        if "credit agreement" in s:
            score += 2
        if re.search(r"no\s+(borrowings|amounts)\s+(were\s+)?outstanding|no\s+outstanding\s+borrowings|no\s+borrowings\s+under", s):
            score += 5
        if re.search(r"borrowings\s+outstanding|outstanding\s+borrowings", s):
            score += 4
        if re.search(r"available|availability|unused", s):
            score += 2
        if re.search(r"letters?\s+of\s+credit|l/c", s):
            score += 1
        if "term loan" in s and "revolv" not in s:
            score -= 3
        if commitment is not None:
            score += 1
        if drawn is not None:
            score += 1
        if availability is not None:
            score += 1
        return score

    def _classify_revolver_facility(
        snippet: str,
        *,
        doc_name: str = "",
        commitment: Optional[float] = None,
        facility_size: Optional[float] = None,
    ) -> Tuple[str, int]:
        ctx = " ".join([str(doc_name or ""), str(snippet or "")]).lower()
        main_hits = 0
        commodity_hits = 0
        other_hits = 0
        main_terms = [
            "working capital",
            "agribusiness",
            "committed",
            "borrowing base",
            "asset-based",
            "asset based",
            "syndicated",
            "revolving credit facility",
            "credit agreement",
            "availability",
            "letters of credit",
        ]
        commodity_terms = [
            "commodity management",
            "margin facility",
            "margin requirements",
            "margin line",
            "hedging",
            "hedge",
            "futures",
            "broker",
            "exchange margin",
            "commodity positions",
            "risk management",
            "clearing",
        ]
        other_terms = [
            "notes payable",
            "other borrowings",
            "short-term borrowings",
            "short term borrowings",
        ]
        main_hits += sum(1 for term in main_terms if term in ctx)
        commodity_hits += sum(1 for term in commodity_terms if term in ctx)
        other_hits += sum(1 for term in other_terms if term in ctx)
        fac_amt = facility_size if facility_size is not None else commitment
        if fac_amt is not None and float(fac_amt) >= 100_000_000 and commodity_hits == 0:
            main_hits += 2
        if fac_amt is not None and float(fac_amt) <= 75_000_000 and commodity_hits > 0:
            commodity_hits += 1
        if commodity_hits >= max(main_hits + 1, 2):
            return "commodity_margin", 10
        if other_hits > max(main_hits, commodity_hits):
            return "other_short_term", 5
        if main_hits > 0 or (revolver_re.search(ctx) and fac_amt is not None and float(fac_amt) >= 100_000_000):
            return "main_committed", 30
        return "unknown", 20

    def _extract_revolver_table(df: pd.DataFrame, scale: float) -> Optional[Dict[str, Any]]:
        if df is None or df.empty:
            return None
        t = df.copy().fillna("")
        header_idx = None
        for i in range(min(3, len(t))):
            row_txt = " ".join([str(x).lower() for x in t.iloc[i].tolist()])
            if any(k in row_txt for k in ["commitment", "outstanding", "available", "availability", "letters of credit", "l/c", "borrowings"]):
                header_idx = i
                break
        if header_idx is not None:
            headers = [str(x).strip().lower() for x in t.iloc[header_idx].tolist()]
            data = t.iloc[header_idx + 1 :].copy()
            data.columns = headers
        else:
            data = t.copy()
            data.columns = [f"col{i}" for i in range(data.shape[1])]

        def _col(keys: List[str]) -> Optional[str]:
            for c in data.columns:
                lc = str(c).lower()
                if any(k in lc for k in keys):
                    return c
            return None

        commit_col = _col(["commitment"])
        out_col = _col(["outstanding", "borrowings"])
        lc_col = _col(["letter", "l/c"])
        avail_col = _col(["available", "availability", "unused"])
        commit_header = str(commit_col).lower() if commit_col is not None else ""

        for _, row in data.iterrows():
            row_text = " ".join([str(x) for x in row.tolist()])
            if "revolv" not in row_text.lower():
                continue

            def _val(col: Optional[str]) -> Optional[float]:
                if not col or col not in row:
                    return None
                v = coerce_number(row[col])
                if v is None:
                    return None
                return float(v) * scale

            commit = _val(commit_col)
            drawn = _val(out_col)
            lc = _val(lc_col)
            avail = _val(avail_col)
            if all(v is None for v in [commit, drawn, lc, avail]):
                continue
            facility = None
            facility_source_type = "missing"
            if commit is not None and "commitment" in commit_header:
                # Distinguish facility size vs borrowing capacity.
                cap_like = any(
                    k in commit_header
                    for k in ["borrowing capacity", "capacity", "available to be borrowed", "permitted borrowings"]
                )
                if not cap_like:
                    facility = commit
                    facility_source_type = "table"
            return {
                "commitment": commit,
                "facility_size": facility,
                "drawn": drawn,
                "lc": lc,
                "availability": avail,
                "facility_source_type": facility_source_type,
                "row_text": row_text,
            }
        return None

    filings_scanned = 0
    min_q = None
    if lookback_years is not None and lookback_years > 0:
        min_q = (datetime.utcnow().date() - timedelta(days=365 * lookback_years))
    cache_path, doc_cache = _load_revolver_doc_cache()
    cache_dirty = False
    for batch in _iter_submission_batches(sec, submissions):
        forms = batch.get("form", [])
        accns = batch.get("accessionNumber", [])
        report_dates = batch.get("reportDate", []) or []
        filing_dates = batch.get("filingDate", []) or []
        primary_docs = batch.get("primaryDocument", []) or []
        n = min(len(forms), len(accns))
        for i in range(n):
            form = forms[i]
            if form not in ("10-Q", "10-Q/A", "10-K", "10-K/A", "8-K", "8-K/A"):
                continue
            accn = accns[i]
            doc = primary_docs[i] if i < len(primary_docs) else None
            rep = report_dates[i] if i < len(report_dates) else None
            fdate = filing_dates[i] if i < len(filing_dates) else None
            q_end = parse_date(rep) or parse_date(fdate)
            if not _is_quarter_end(q_end):
                if form.startswith("8-K"):
                    # 8-K earnings exhibits are typically filed after quarter-end; anchor to latest closed quarter.
                    q_end = _coerce_prev_quarter_end(q_end)
                else:
                    q_end = _coerce_prev_quarter_end(q_end)
            if min_q is not None and q_end and q_end < min_q:
                continue
            accn_nd = normalize_accession(accn)

            cand_docs: List[Tuple[str, str]] = []
            seen_docs: set = set()
            if doc:
                ok, reason = _doc_filter_reason(doc, is_primary=True)
                if ok:
                    cand_docs.append((doc, reason))
                    seen_docs.add(doc.lower())
                else:
                    print(f"[revolver] SKIP doc={doc} reason={reason}", flush=True)
            try:
                idx = sec.accession_index_json(cik_int, accn_nd)
            except Exception:
                idx = {}
            items = idx.get("directory", {}).get("item", []) if isinstance(idx, dict) else []
            for it in items:
                name = it.get("name")
                if not name:
                    continue
                low = name.lower()
                if low in seen_docs:
                    continue
                ok, reason = _doc_filter_reason(name, is_primary=False)
                if not ok:
                    print(f"[revolver] SKIP doc={name} reason={reason}", flush=True)
                    continue
                cand_docs.append((name, reason))
                seen_docs.add(low)
            cand_docs = cand_docs[:REVOLVER_DOC_MAX_PER_FILING]
            print(
                f"[revolver] filing form={form} accn={accn_nd} q={q_end} docs={len(cand_docs)}",
                flush=True,
            )
            doc_scanned = 0

            for doc_name, doc_reason in cand_docs:
                if doc_scanned >= REVOLVER_DOC_MAX_PER_FILING:
                    break
                t_doc = time.perf_counter()
                print(f"[revolver] {accn_nd} downloading {doc_name}...", flush=True)
                t_dl0 = time.perf_counter()
                try:
                    html_bytes = sec.download_document(cik_int, accn_nd, doc_name)
                except Exception as ex:
                    print(f"[revolver] {accn_nd} {doc_name} download failed: {type(ex).__name__}", flush=True)
                    continue
                t_dl1 = time.perf_counter()
                html_text = html_bytes.decode("utf-8", errors="ignore")
                doc_sha1 = hashlib.sha1(html_bytes).hexdigest()
                cache_key = f"{accn_nd}:{doc_name.lower()}"
                cache_entry = doc_cache.get(cache_key)
                if (
                    isinstance(cache_entry, dict)
                    and cache_entry.get("sha1") == doc_sha1
                    and isinstance(cache_entry.get("result_rows"), list)
                ):
                    print(f"[revolver] cache hit accn={accn_nd} doc={doc_name}", flush=True)
                    for rr in cache_entry.get("result_rows", []):
                        if isinstance(rr, dict):
                            restored = dict(rr)
                            for dcol in ("quarter", "filed", "report_date"):
                                if dcol in restored:
                                    restored[dcol] = parse_date(restored.get(dcol))
                            rows.append(restored)
                    doc_scanned += 1
                    continue
                t_pre0 = time.perf_counter()
                pre_text = _quick_scan_text(html_text)
                pre_scan_hit = _contains_revolver_keyword(pre_text)
                t_pre1 = time.perf_counter()
                if not pre_scan_hit:
                    print(f"[revolver] SKIP doc={doc_name} reason=no-keyword-prescan", flush=True)
                    doc_cache[cache_key] = {
                        "sha1": doc_sha1,
                        "pre_scan_hit": False,
                        "candidate_indexes": [],
                        "result_rows": [],
                    }
                    cache_dirty = True
                    continue
                text = pre_text
                if not text:
                    print(f"[revolver] {accn_nd} {doc_name} empty text after strip_html; skip", flush=True)
                    continue
                scale = _detect_scale_from_text(text)

                t_tbl0 = time.perf_counter()
                print(
                    f"[revolver] {accn_nd} {doc_name} bytes={len(html_bytes)} parsing tables...",
                    flush=True,
                )
                table_blocks = _extract_table_blocks(html_text)
                candidate_indexes: List[int] = []
                candidate_blocks: List[str] = []
                for ti, tb in enumerate(table_blocks):
                    if len(candidate_blocks) >= REVOLVER_TABLE_MAX_CANDIDATES:
                        break
                    table_text = _quick_scan_text(tb)
                    if _table_is_revolver_candidate(table_text):
                        candidate_indexes.append(ti)
                        candidate_blocks.append(tb)
                t_tbl1 = time.perf_counter()
                t_parse0 = time.perf_counter()
                tables: List[pd.DataFrame] = []
                for tb in candidate_blocks:
                    try:
                        parsed_tbls = pd.read_html(io.StringIO(tb))
                    except Exception:
                        parsed_tbls = []
                    if not parsed_tbls:
                        try:
                            parsed_tbls = read_html_tables_any(tb.encode("utf-8", errors="ignore"))
                        except Exception:
                            parsed_tbls = []
                    if parsed_tbls:
                        tables.extend(parsed_tbls)
                t_parse1 = time.perf_counter()
                print(
                    f"[revolver] {accn_nd} {doc_name} tables={len(table_blocks)} candidates={len(candidate_blocks)} parsed={len(tables)} done",
                    flush=True,
                )
                doc_rows: List[Dict[str, Any]] = []
                for t in tables:
                    parsed = _extract_revolver_table(t, scale)
                    if not parsed:
                        continue
                    row_text = parsed.get("row_text") or ""
                    facility_class, facility_priority = _classify_revolver_facility(
                        row_text,
                        doc_name=doc_name,
                        commitment=parsed.get("commitment"),
                        facility_size=parsed.get("facility_size"),
                    )
                    rows.append({
                        "quarter": q_end,
                        "accn": accn,
                        "form": form,
                        "filed": parse_date(fdate),
                        "report_date": parse_date(rep),
                        "doc": doc_name,
                        "revolver_commitment": parsed.get("commitment"),
                        "revolver_facility_size": parsed.get("facility_size"),
                        "revolver_drawn": parsed.get("drawn"),
                        "revolver_lc": parsed.get("lc"),
                        "revolver_availability": parsed.get("availability"),
                        "commitment_source_type": "table" if parsed.get("commitment") is not None else "missing",
                        "facility_source_type": parsed.get("facility_source_type") or ("table" if parsed.get("facility_size") is not None else "missing"),
                        "drawn_source_type": "table" if parsed.get("drawn") is not None else "missing",
                        "lc_source_type": "table" if parsed.get("lc") is not None else "missing",
                        "availability_source_type": "table" if parsed.get("availability") is not None else "missing",
                        "commitment_snippet": row_text if parsed.get("commitment") is not None else None,
                        "drawn_snippet": row_text if parsed.get("drawn") is not None else None,
                        "lc_snippet": row_text if parsed.get("lc") is not None else None,
                        "availability_snippet": row_text if parsed.get("availability") is not None else None,
                        "source_type": "table",
                        "source_snippet": row_text,
                        "snippet": row_text,
                        "source_class": "filing_table",
                        "method": "revolver_table",
                        "qa_severity": "warn",
                        "score": 100,
                        "facility_class": facility_class,
                        "facility_priority": facility_priority,
                    })

                doc_commit = None
                doc_commit_snip = None
                doc_commit_pri = -1
                doc_facility = None
                doc_facility_snip = None
                doc_lc = None
                doc_lc_snip = None
                for pat in (agg_re, rev_re1, rev_re2, perm_re, capacity_re):
                    for m_doc in pat.finditer(text):
                        if not _near_revolver(text, m_doc.start(), window=120):
                            continue
                        ctx = _snippet(text, m_doc.start(), m_doc.end(), 120)
                        if pat is agg_re and not revolver_re.search(ctx):
                            continue
                        if pat not in (rev_re1, rev_re2) and not _amount_ok(ctx, m_doc.group(0)):
                            continue
                        val = _scale_num(m_doc.group(1), m_doc.group(2) or "", ctx)
                        if val is None:
                            continue
                        if pat in (agg_re, perm_re, capacity_re):
                            if pat is perm_re or pat is agg_re:
                                pri = 2
                            else:
                                pri = 1
                            if pri >= doc_commit_pri:
                                doc_commit = float(val)
                                doc_commit_snip = ctx
                                doc_commit_pri = pri
                        else:
                            doc_facility = float(val)
                            doc_facility_snip = ctx
                # Capture outstanding revolver L/C from full-document text when present.
                lc_doc_patterns = [
                    re.compile(
                        r"(?:approximately|about)?\s*\$?\s*([0-9]{1,3}(?:,[0-9]{3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)\s*"
                        r"(million|billion)?\s+outstanding\s+letters?\s+of\s+credit",
                        re.I,
                    ),
                    re.compile(
                        r"outstanding\s+letters?\s+of\s+credit[^0-9]{0,90}"
                        r"(?:approximately|about)?\s*\$?\s*([0-9]{1,3}(?:,[0-9]{3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)\s*(million|billion)?",
                        re.I,
                    ),
                ]
                for pat_lc in lc_doc_patterns:
                    if doc_lc is not None:
                        break
                    for m_doc_lc in pat_lc.finditer(text):
                        ctx_lc = _snippet(text, m_doc_lc.start(), m_doc_lc.end(), 320)
                        forward_ctx = text[m_doc_lc.start() : min(len(text), m_doc_lc.start() + 900)]
                        if not re.search(
                            r"revolv|credit\s+facility|reduce\s+the\s+amount\s+we\s+can\s+borrow|borrow\s+under",
                            forward_ctx,
                            re.I,
                        ):
                            continue
                        lc_val = _scale_num(m_doc_lc.group(1), m_doc_lc.group(2) or "", ctx_lc, m_doc_lc.group(0))
                        if lc_val is not None:
                            doc_lc = float(lc_val)
                            doc_lc_snip = ctx_lc
                            break
                if revolver_re.search(text):
                    m0 = revolver_re.search(text)
                    snip0 = _snippet(text, m0.start(), m0.end())
                    future0 = _is_future_event(snip0, q_end)
                    doc_commit2, doc_commit_method, doc_facility2, _doc_fac_m = _find_commitment_with_method(snip0)
                    if doc_commit2 is not None:
                        if future0 and doc_commit_method not in {"change_phrase", "permitted_borrowings", "aggregate_commitments", "borrowing_capacity"}:
                            pass
                        else:
                            if doc_commit_method in {"change_phrase"}:
                                doc_commit = float(doc_commit2)
                                doc_commit_snip = snip0
                                doc_commit_pri = 3
                            elif doc_commit_method in {"permitted_borrowings", "aggregate_commitments", "borrowing_capacity"}:
                                if doc_commit_pri < 2:
                                    doc_commit = float(doc_commit2)
                                    doc_commit_snip = snip0
                                    doc_commit_pri = 2
                            elif doc_commit_method in {"provides_for"}:
                                if doc_commit_pri < 1:
                                    doc_commit = float(doc_commit2)
                                    doc_commit_snip = snip0
                                    doc_commit_pri = 1
                    if doc_facility2 is not None and not future0:
                        doc_facility = float(doc_facility2)
                        doc_facility_snip = snip0

                candidates: List[Dict[str, Any]] = []
                for m in revolver_re.finditer(text):
                    snip = _snippet(text, m.start(), m.end())
                    commitment, commit_method, facility_size, facility_method = _find_commitment_with_method(snip)
                    asof = _asof_date_in_text(snip)
                    q_end_snip = q_end
                    if asof:
                        q_end_snip = asof if _is_quarter_end(asof) else _coerce_prev_quarter_end(asof)
                    future_event = _is_future_event(snip, q_end_snip)
                    if commit_method in {"change_phrase", "permitted_borrowings", "aggregate_commitments", "borrowing_capacity"}:
                        future_event = False
                    # Do not guess facility size from unrelated numbers in the snippet.
                    # If capacity not mentioned separately, use facility size as capacity
                    if commitment is None and facility_size is not None:
                        commitment = facility_size
                        commit_method = commit_method or "facility_default"
                    drawn = None
                    lc_amt = None
                    if re.search(r"no\s+(borrowings|amounts)\s+(were\s+)?outstanding|no\s+outstanding\s+borrowings|no\s+borrowings\s+under", snip, re.I):
                        drawn = 0.0
                    else:
                        m2 = re.search(r"borrowings\s+outstanding[^0-9]{0,80}([0-9]{1,3}(?:,[0-9]{3})+(?:\\.\\d+)?|[0-9]+(?:\\.\\d+)?)(?:\s*(million|billion))?", snip, re.I)
                        if m2:
                            drawn = _scale_num(m2.group(1), m2.group(2) or "", snip)
                    if drawn is None:
                        m3 = re.search(r"borrowings\s+outstanding\s+under\s+the\s+revolving\s+credit\s+facility[^0-9]{0,80}([0-9]{1,3}(?:,[0-9]{3})+(?:\\.\\d+)?|[0-9]+(?:\\.\\d+)?)(?:\s*(million|billion))?", snip, re.I)
                        if m3:
                            drawn = _scale_num(m3.group(1), m3.group(2) or "", snip)
                    if drawn is None and "revolving credit facility" in snip.lower():
                        m4 = re.search(r"revolving\s+credit\s+facility[^.]{0,160}borrowings[^0-9]{0,60}([0-9]{1,3}(?:,[0-9]{3})+(?:\\.\\d+)?|[0-9]+(?:\\.\\d+)?)(?:\s*(million|billion))?", snip, re.I)
                        if m4:
                            drawn = _scale_num(m4.group(1), m4.group(2) or "", snip)
                    m_lc = re.search(
                        r"outstanding\s+letters?\s+of\s+credit[^0-9]{0,80}"
                        r"([0-9]{1,3}(?:,[0-9]{3})+(?:\\.\\d+)?|[0-9]+(?:\\.\\d+)?)(?:\s*(million|billion))?",
                        snip,
                        re.I,
                    )
                    if m_lc:
                        lc_amt = _scale_num(m_lc.group(1), m_lc.group(2) or "", snip, m_lc.group(0))
                    if lc_amt is None:
                        m_lc_before = re.search(
                            r"(?:approximately|about)?\s*\$?\s*([0-9]{1,3}(?:,[0-9]{3})+(?:\\.\\d+)?|[0-9]+(?:\\.\\d+)?)\s*"
                            r"(million|billion)?\s+outstanding\s+letters?\s+of\s+credit",
                            snip,
                            re.I,
                        )
                        if m_lc_before:
                            lc_amt = _scale_num(m_lc_before.group(1), m_lc_before.group(2) or "", snip, m_lc_before.group(0))
                    if lc_amt is None and doc_lc is not None:
                        lc_amt = float(doc_lc)
                    if lc_amt is None and "letters of credit" in text.lower():
                        # Try to capture LC from nearby lines in full text
                        for m_lc2 in re.finditer(r"(?:outstanding\s+letters?\s+of\s+credit|letters?\s+of\s+credit[^.\n]{0,120}outstanding)[^.\n]{0,200}", text, re.I):
                            ctx_lc = _snippet(text, m_lc2.start(), m_lc2.end(), 200)
                            if not re.search(r"revolv|credit\s+facility|reduce\s+the\s+amount\s+we\s+can\s+borrow|borrow\s+under", ctx_lc, re.I):
                                continue
                            m_lc3 = re.search(
                                r"(?:approximately|about)?\s*\$?\s*([0-9]{1,3}(?:,[0-9]{3})+(?:\\.\\d+)?|[0-9]+(?:\\.\\d+)?)\s*(million|billion)?",
                                ctx_lc,
                                re.I,
                            )
                            if m_lc3:
                                lc_amt = _scale_num(m_lc3.group(1), m_lc3.group(2) or "", ctx_lc, m_lc3.group(0))
                                if lc_amt is not None:
                                    break
                    if lc_amt is None and doc_lc is not None:
                        lc_amt = float(doc_lc)
                    if drawn is None and commitment is not None:
                        if re.search(
                            r"no\s+(borrowings|amounts)\s+(were\s+)?outstanding|no\s+outstanding\s+borrowings|no\s+borrowings\s+under",
                            text,
                            re.I,
                        ):
                            drawn = 0.0
                    availability = None
                    availability_derived = False
                    m_av = re.search(r"available[^0-9]{0,50}\$?\s*([0-9]{1,3}(?:,[0-9]{3})+(?:\.\d+)?|[0-9]+(?:\.\d+)?)(?:\s*(million|billion))?", snip, re.I)
                    if m_av:
                        availability = _scale_num(m_av.group(1), m_av.group(2) or "", snip)
                    if availability is None and commitment is not None and drawn is not None and lc_amt is not None:
                        availability = commitment - drawn - lc_amt
                        availability_derived = True
                    score = _score_snip(snip, commitment, drawn, availability)
                    if future_event:
                        score -= 5
                    facility_class, facility_priority = _classify_revolver_facility(
                        snip,
                        doc_name=doc_name,
                        commitment=commitment,
                        facility_size=facility_size,
                    )
                    commitment_snip = snip if commitment is not None else None
                    drawn_snip = snip if drawn is not None else None
                    lc_snip = doc_lc_snip if (lc_amt is not None and doc_lc_snip) else (snip if lc_amt is not None else None)
                    availability_snip = snip if availability is not None else None
                    candidates.append({
                        "quarter": q_end_snip,
                        "accn": accn,
                        "form": form,
                        "filed": parse_date(fdate),
                        "report_date": parse_date(rep),
                        "doc": doc_name,
                        "revolver_commitment": commitment,
                        "revolver_facility_size": facility_size,
                        "revolver_drawn": drawn,
                        "revolver_lc": lc_amt,
                        "revolver_availability": availability,
                        "commitment_source_type": "text" if commitment is not None else "missing",
                        "facility_source_type": "text" if facility_size is not None else "missing",
                        "drawn_source_type": "text" if drawn is not None else "missing",
                        "lc_source_type": "text" if lc_amt is not None else "missing",
                        "availability_source_type": ("derived" if availability_derived else ("text" if availability is not None else "missing")),
                        "commitment_snippet": commitment_snip,
                        "drawn_snippet": drawn_snip,
                        "lc_snippet": lc_snip,
                        "availability_snippet": availability_snip,
                        "source_type": "text",
                        "source_snippet": snip,
                        "snippet": snip,
                        "source_class": "filing_text",
                        "method": f"revolver_scan:{commit_method or (facility_method or 'unknown')}",
                        "qa_severity": "warn",
                        "score": score,
                        "future_event": future_event,
                        "facility_class": facility_class,
                        "facility_priority": facility_priority,
                    })
                if candidates:
                    # Prefer non-future change phrases / permitted borrowings where available
                    preferred = [c for c in candidates if not c.get("future_event") and str(c.get("method", "")).find("change_phrase") >= 0]
                    if not preferred:
                        preferred = [c for c in candidates if not c.get("future_event") and str(c.get("method", "")).find("permitted_borrowings") >= 0]
                    if not preferred:
                        preferred = [c for c in candidates if not c.get("future_event")]
                    if not preferred:
                        preferred = candidates
                    preferred.sort(
                        key=lambda r: (
                            r.get("facility_priority", 0),
                            r.get("revolver_commitment") is not None,
                            float(r.get("revolver_commitment") or r.get("revolver_facility_size") or 0.0),
                            r.get("revolver_drawn") is not None,
                            r.get("revolver_facility_size") is not None,
                            r.get("score", 0),
                        ),
                        reverse=True,
                    )
                    doc_rows.append(preferred[0])
                    found = True
                else:
                    found = False
                has_commit_candidate = any(c.get("revolver_commitment") is not None for c in candidates)
                if doc_commit is not None and doc_commit_snip and (not has_commit_candidate):
                    facility_class, facility_priority = _classify_revolver_facility(
                        doc_commit_snip,
                        doc_name=doc_name,
                        commitment=doc_commit,
                        facility_size=doc_facility,
                    )
                    doc_rows.append({
                        "quarter": q_end,
                        "accn": accn,
                        "form": form,
                        "filed": parse_date(fdate),
                        "report_date": parse_date(rep),
                        "doc": doc_name,
                        "revolver_commitment": doc_commit,
                        "revolver_facility_size": doc_facility,
                        "revolver_drawn": None,
                        "revolver_lc": None,
                        "revolver_availability": None,
                        "commitment_source_type": "text",
                        "facility_source_type": "text" if doc_facility is not None else "missing",
                        "drawn_source_type": "missing",
                        "lc_source_type": "missing",
                        "availability_source_type": "missing",
                        "commitment_snippet": doc_commit_snip,
                        "drawn_snippet": None,
                        "lc_snippet": None,
                        "availability_snippet": None,
                        "source_type": "text",
                        "source_snippet": doc_commit_snip,
                        "snippet": doc_commit_snip,
                        "source_class": "filing_text",
                        "method": "doc_commit",
                        "qa_severity": "warn",
                        "score": 50,
                        "facility_class": facility_class,
                        "facility_priority": facility_priority,
                    })
                t_post1 = time.perf_counter()
                rows.extend(doc_rows)
                cache_rows = [{k: _json_safe(v) for k, v in r.items()} for r in doc_rows]
                doc_cache[cache_key] = {
                    "sha1": doc_sha1,
                    "pre_scan_hit": True,
                    "candidate_indexes": candidate_indexes,
                    "result_rows": cache_rows,
                }
                cache_dirty = True
                print(
                    "[revolver] "
                    f"{accn_nd} {doc_name} timing "
                    f"download={t_dl1 - t_dl0:.2f}s "
                    f"pre_scan={t_pre1 - t_pre0:.2f}s "
                    f"extract_tables={t_tbl1 - t_tbl0:.2f}s "
                    f"parse_candidates={t_parse1 - t_parse0:.2f}s "
                    f"postprocess={t_post1 - t_parse1:.2f}s "
                    f"total={t_post1 - t_doc:.2f}s",
                    flush=True,
                )
                doc_scanned += 1
            filings_scanned += 1
            if filings_scanned >= max_docs:
                break
        if filings_scanned >= max_docs:
            break
    if cache_dirty:
        _dump_revolver_doc_cache(cache_path, doc_cache)
    return pd.DataFrame(rows)


def build_revolver_history(
    revolver_df: pd.DataFrame,
    hist: pd.DataFrame,
    capacity_map: Optional[Dict[pd.Timestamp, float]] = None,
    capacity_meta: Optional[Dict[pd.Timestamp, Dict[str, Any]]] = None,
    max_quarters: int = 20,
) -> pd.DataFrame:
    if (revolver_df is None or revolver_df.empty) and not capacity_map:
        return pd.DataFrame()
    if hist is None or hist.empty:
        return pd.DataFrame()
    hq = pd.to_datetime(hist["quarter"], errors="coerce")
    qs = sorted(hq.dropna().unique())
    if len(qs) > max_quarters:
        qs = qs[-max_quarters:]
    rev = revolver_df.copy() if revolver_df is not None else pd.DataFrame()
    if not rev.empty:
        rev["quarter"] = pd.to_datetime(rev["quarter"], errors="coerce")
        rev = rev[rev["quarter"].notna()]

    source_rank = {"xbrl": 0, "table": 1, "text": 2, "derived": 3, "missing": 4}
    if not rev.empty:
        rev["source_type"] = rev.get("source_type").fillna("text")
        rev["source_rank"] = rev["source_type"].map(source_rank).fillna(9)
        rev["score"] = pd.to_numeric(rev.get("score"), errors="coerce").fillna(0)
        rev["has_commit"] = pd.to_numeric(rev.get("revolver_commitment"), errors="coerce").notna()
        rev["has_drawn"] = pd.to_numeric(rev.get("revolver_drawn"), errors="coerce").notna()
        rev["has_facility"] = pd.to_numeric(rev.get("revolver_facility_size"), errors="coerce").notna()
        rev["facility_amount"] = (
            pd.to_numeric(rev.get("revolver_commitment"), errors="coerce")
            .fillna(pd.to_numeric(rev.get("revolver_facility_size"), errors="coerce"))
            .fillna(0.0)
        )
        rev["facility_priority"] = pd.to_numeric(rev.get("facility_priority"), errors="coerce").fillna(0)
        rev["facility_class"] = rev.get("facility_class").fillna("unknown")
        rev = rev.sort_values(
            ["quarter", "facility_priority", "source_rank", "has_commit", "facility_amount", "score", "has_drawn", "has_facility"],
            ascending=[True, False, True, False, False, False, False, False],
        )

    last_commit = None
    last_commit_q = None
    last_facility = None
    last_facility_q = None
    rows = []
    def _none_if_nan(x: Any) -> Optional[float]:
        if x is None:
            return None
        try:
            if pd.isna(x):
                return None
        except Exception:
            pass
        try:
            return float(x)
        except Exception:
            return None

    def _norm_src(val: Optional[str]) -> str:
        if not val:
            return "missing"
        v = str(val).lower()
        if v in ("table", "filing_table"):
            return "table"
        if v in ("text", "xbrl", "derived", "missing", "table"):
            return v
        return "text"

    def _first_nonnull_num(frame: pd.DataFrame, col: str) -> Optional[float]:
        if frame is None or frame.empty or col not in frame.columns:
            return None
        vals = pd.to_numeric(frame[col], errors="coerce")
        idxs = vals[vals.notna()].index.tolist()
        if not idxs:
            return None
        try:
            return float(vals.loc[idxs[0]])
        except Exception:
            return None

    for q in qs:
        q = pd.Timestamp(q)
        sub = rev[rev["quarter"] == q] if not rev.empty and "quarter" in rev.columns else pd.DataFrame()
        if not sub.empty:
            row = sub.iloc[0]
            commit = _none_if_nan(row.get("revolver_commitment"))
            facility = _none_if_nan(row.get("revolver_facility_size"))
            drawn = _none_if_nan(row.get("revolver_drawn"))
            lc = _none_if_nan(row.get("revolver_lc"))
            avail = _none_if_nan(row.get("revolver_availability"))
            source_type = row.get("source_type") or "text"
            snippet = row.get("source_snippet") or row.get("snippet")
            commit_source = _norm_src(row.get("commitment_source_type") or source_type)
            facility_source = _norm_src(row.get("facility_source_type") or source_type)
            drawn_source = _norm_src(row.get("drawn_source_type") or source_type)
            lc_source = _norm_src(row.get("lc_source_type") or source_type)
            avail_source = _norm_src(row.get("availability_source_type") or source_type)
            commit_snip = row.get("commitment_snippet") or snippet
            drawn_snip = row.get("drawn_snippet") or snippet
            lc_snip = row.get("lc_snippet") or snippet
            avail_snip = row.get("availability_snippet") or snippet
            note = row.get("note") or ""
            preferred_class = str(row.get("facility_class") or "unknown").strip().lower() or "unknown"
            sub_same = sub[sub.get("facility_class").fillna("unknown").astype(str).str.lower() == preferred_class].copy() if "facility_class" in sub.columns else sub
            if sub_same.empty:
                sub_same = sub
            # If we have multiple rows for the quarter, only fill commitment if missing
            commit_alt = _first_nonnull_num(sub_same, "revolver_commitment")
            if commit is None and commit_alt is not None:
                commit = commit_alt
                commit_source = "text"
                note = (note + "; " if note else "") + "commitment from other source"
            facility_alt = _first_nonnull_num(sub_same, "revolver_facility_size")
            if facility is None and facility_alt is not None:
                facility = facility_alt
                facility_source = "text"
                note = (note + "; " if note else "") + "facility size from other source"
            if drawn is None:
                drawn_alt = _first_nonnull_num(sub_same, "revolver_drawn")
                if drawn_alt is not None:
                    drawn = drawn_alt
                    drawn_source = "text"
                    note = (note + "; " if note else "") + "drawn from other source"
            if lc is None and "revolver_lc" in sub_same.columns:
                lc_alt = _first_nonnull_num(sub_same, "revolver_lc")
                if lc_alt is not None:
                    if commit is None or (0.0 <= lc_alt <= float(commit) * 1.05):
                        lc = lc_alt
                        lc_source = "text"
                        note = (note + "; " if note else "") + "L/C from other source"
            if avail is None:
                avail_alt = _first_nonnull_num(sub_same, "revolver_availability")
                if avail_alt is not None:
                    if commit is None or (0.0 <= avail_alt <= float(commit) * 1.05):
                        avail = avail_alt
                        avail_source = "text"
                        note = (note + "; " if note else "") + "availability from other source"
        else:
            commit = None
            facility = None
            drawn = None
            lc = None
            avail = None
            source_type = "missing"
            snippet = None
            note = ""
            commit_source = "missing"
            facility_source = "missing"
            drawn_source = "missing"
            lc_source = "missing"
            avail_source = "missing"
            commit_snip = None
            drawn_snip = None
            lc_snip = None
            avail_snip = None
        # XBRL capacity (LineOfCreditFacilityMaximumBorrowingCapacity) as primary commitment if available
        cap_key = pd.Timestamp(q.date())
        cap_val = capacity_map.get(cap_key) if capacity_map else None
        cap_meta = capacity_meta.get(cap_key) if capacity_meta else None
        if cap_val is not None and cap_meta:
            end_d = cap_meta.get("end_d")
            if end_d and end_d > q.date():
                cap_val = None
                note = (note + "; " if note else "") + "xbrl capacity after quarter_end"
        if cap_val is not None:
            cap_val = _none_if_nan(cap_val)
            if cap_val is not None and not (100e6 <= float(cap_val) <= 800e6):
                scaled_val = None
                scaled_note = ""
                if float(cap_val) < 100e6:
                    cand1 = float(cap_val) * 1_000.0
                    cand2 = float(cap_val) * 1_000_000.0
                    if 100e6 <= cand1 <= 800e6:
                        scaled_val = cand1
                        scaled_note = "x1000"
                    elif 100e6 <= cand2 <= 800e6:
                        scaled_val = cand2
                        scaled_note = "x1e6"
                elif float(cap_val) > 800e6:
                    cand1 = float(cap_val) / 1_000.0
                    cand2 = float(cap_val) / 1_000_000.0
                    if 100e6 <= cand1 <= 800e6:
                        scaled_val = cand1
                        scaled_note = "/1000"
                    elif 100e6 <= cand2 <= 800e6:
                        scaled_val = cand2
                        scaled_note = "/1e6"
                if scaled_val is not None:
                    cap_val = scaled_val
                    if commit is None:
                        note = (note + "; " if note else "") + f"xbrl capacity scaled {scaled_note}"
                else:
                    if commit is None:
                        note = (note + "; " if note else "") + "xbrl capacity out of range"
                    cap_val = None
        if commit is None and cap_val is not None:
            commit = cap_val
            commit_source = "xbrl"
            if source_type == "missing":
                source_type = "xbrl"
            note = (note + "; " if note else "") + "capacity from XBRL"
            if cap_meta:
                accn = cap_meta.get("accn")
                end_d = cap_meta.get("end_d")
                if accn:
                    note = (note + "; " if note else "") + f"xbrl accn {accn}"
                if end_d and end_d != q.date():
                    note = (note + "; " if note else "") + "xbrl end_d != quarter_end"
            commit_snip = commit_snip or "xbrl capacity"
        if commit is not None:
            last_commit = commit
            last_commit_q = q
        if commit is None and last_commit is not None and last_commit_q is not None:
            if (q - last_commit_q).days <= 1860:
                commit = last_commit
                commit_source = "derived"
                source_type = "derived"
                note = (note + "; " if note else "") + "commitment carry-forward"
        if facility is not None:
            last_facility = facility
            last_facility_q = q
        if facility is None and commit is None and last_facility is not None and last_facility_q is not None:
            if (q - last_facility_q).days <= 1860:
                facility = last_facility
                facility_source = "derived"
                note = (note + "; " if note else "") + "facility size carry-forward"
        if facility is None and commit is not None:
            facility = float(commit)
            facility_source = "derived"
            note = (note + "; " if note else "") + "facility set to capacity"
        if facility is not None and commit is not None and facility < commit:
            facility = commit
            facility_source = "derived"
            note = (note + "; " if note else "") + "facility < capacity; set to capacity"
        if avail is None and commit is not None and drawn is not None and lc is not None and source_type in ("table", "text", "xbrl"):
            avail = float(commit) - float(drawn) - float(lc)
            avail_source = "derived"
            note = (note + "; " if note else "") + "availability derived"
        elif avail is None and commit is not None and drawn is not None and lc is None:
            avail = float(commit) - float(drawn)
            avail_source = "derived"
            note = (note + "; " if note else "") + "lc_missing"
        if avail is not None and commit is not None:
            # Keep availability internally consistent with capacity after all carry-forward steps.
            cap_val = float(commit)
            if float(avail) > cap_val + 1e-6:
                if drawn is not None:
                    recomputed = cap_val - float(drawn) - (float(lc) if lc is not None else 0.0)
                    avail = max(recomputed, 0.0)
                    note = (note + "; " if note else "") + "availability > capacity; recomputed"
                else:
                    avail = cap_val
                    note = (note + "; " if note else "") + "availability > capacity; capped"
                avail_source = "derived"
        utilization = None
        if commit is not None and drawn is not None and commit != 0:
            utilization = float(drawn) / float(commit)
        if avail is not None and float(avail) < 0:
            note = (note + "; " if note else "") + "qa: availability < 0"
        if commit is not None and drawn is not None and float(drawn) > float(commit):
            note = (note + "; " if note else "") + "qa: drawn > commitment"
        if utilization is not None and float(utilization) > 1.0:
            note = (note + "; " if note else "") + "qa: utilization > 1.0"
        rows.append({
            "quarter": q.date(),
            "revolver_commitment": commit,
            "revolver_facility_size": facility,
            "revolver_drawn": drawn,
            "revolver_letters_of_credit": lc,
            "revolver_availability": avail,
            "revolver_utilization": utilization,
            "commitment_source_type": commit_source,
            "facility_source_type": facility_source,
            "drawn_source_type": drawn_source,
            "lc_source_type": lc_source,
            "availability_source_type": avail_source,
            "commitment_snippet": commit_snip,
            "drawn_snippet": drawn_snip,
            "lc_snippet": lc_snip,
            "availability_snippet": avail_snip,
            "source_type": source_type,
            "source_snippet": snippet,
            "note": note,
        })
    return pd.DataFrame(rows)


def build_local_main_revolver_history(
    base_dir: Path,
    *,
    ticker: Optional[str] = None,
    cache_root: Optional[Path] = None,
    rebuild_doc_text_cache: bool = False,
    quiet_pdf_warnings: bool = True,
) -> pd.DataFrame:
    fs_dir = Path(base_dir) / "financial_statement"

    def _parse_quarter_from_name(name: str) -> Optional[dt.date]:
        nm = str(name or "")
        m_date = re.search(r"\b(20\d{2})[-_](\d{2})[-_](\d{2})\b", nm)
        if m_date:
            try:
                return dt.date(int(m_date.group(1)), int(m_date.group(2)), int(m_date.group(3)))
            except Exception:
                pass
        m_fy = re.search(r"(?<!\d)fy[-_ ]?(20\d{2})(?!\d)", nm, re.I)
        if m_fy:
            try:
                return dt.date(int(m_fy.group(1)), 12, 31)
            except Exception:
                pass
        m_q = re.search(r"(?<!\d)Q([1-4])[-_ ]?(20\d{2})(?!\d)", nm, re.I)
        if not m_q:
            m_q = re.search(r"(?<!\d)(20\d{2})[-_ ]?Q([1-4])(?!\d)", nm, re.I)
            if m_q:
                year = int(m_q.group(1))
                qnum = int(m_q.group(2))
            else:
                return None
        else:
            qnum = int(m_q.group(1))
            year = int(m_q.group(2))
        try:
            return pd.Period(f"{year}Q{qnum}", freq="Q").end_time.date()
        except Exception:
            return None

    def _read_text(path_in: Path) -> str:
        try:
            suf = path_in.suffix.lower()
            if suf == ".txt":
                return path_in.read_text(encoding="utf-8", errors="ignore")
            if suf in {".htm", ".html"}:
                return strip_html(path_in.read_text(encoding="utf-8", errors="ignore"))
            if suf == ".pdf":
                return extract_pdf_text_cached(
                    path_in,
                    cache_root=(cache_root or base_dir),
                    rebuild_cache=rebuild_doc_text_cache,
                    quiet_pdf_warnings=quiet_pdf_warnings,
                )
        except Exception:
            return ""
        return ""

    def _scale_amount(num_str: str, unit: str = "", *, default_millions: bool = False) -> Optional[float]:
        try:
            val = float(str(num_str or "").replace(",", ""))
        except Exception:
            return None
        unit_l = str(unit or "").strip().lower()
        if unit_l.startswith("b"):
            return val * 1_000_000_000.0
        if unit_l.startswith("m"):
            return val * 1_000_000.0
        if default_millions:
            return val * 1_000_000.0
        return val

    def _classify_local_revolver_facility(
        snippet: str,
        *,
        doc_name: str = "",
        commitment: Optional[float] = None,
        facility_size: Optional[float] = None,
    ) -> Tuple[str, int]:
        ctx = " ".join([str(doc_name or ""), str(snippet or "")]).lower()
        main_hits = 0
        commodity_hits = 0
        other_hits = 0
        main_terms = [
            "working capital",
            "agribusiness",
            "committed",
            "borrowing base",
            "asset-based",
            "asset based",
            "syndicated",
            "revolver",
            "revolving credit facility",
            "credit agreement",
            "availability",
            "letters of credit",
            "green plains finance company",
            "green plains grain",
            "green plains trade",
        ]
        commodity_terms = [
            "commodity management",
            "margin facility",
            "margin line",
            "hedging",
            "hedge",
            "futures",
            "exchange margin",
            "commodity positions",
            "risk management",
            "clearing",
        ]
        other_terms = [
            "notes payable",
            "other borrowings",
            "short-term borrowings",
            "short term borrowings",
            "ancora",
        ]
        main_hits += sum(1 for term in main_terms if term in ctx)
        commodity_hits += sum(1 for term in commodity_terms if term in ctx)
        other_hits += sum(1 for term in other_terms if term in ctx)
        fac_amt = facility_size if facility_size is not None else commitment
        if fac_amt is not None and float(fac_amt) >= 100_000_000 and commodity_hits == 0:
            main_hits += 2
        if fac_amt is not None and float(fac_amt) <= 75_000_000 and commodity_hits > 0:
            commodity_hits += 1
        if commodity_hits >= max(main_hits + 1, 2):
            return "commodity_margin", 10
        if other_hits > max(main_hits, commodity_hits):
            return "other_short_term", 5
        if main_hits > 0 or (fac_amt is not None and float(fac_amt) >= 100_000_000):
            return "main_committed", 30
        return "unknown", 20

    def _parse_doc(path_in: Path) -> Optional[Dict[str, Any]]:
        raw_txt = _read_text(path_in)
        if not raw_txt:
            return None
        q_end = _parse_quarter_from_name(path_in.name) or infer_quarter_end_from_text(raw_txt)
        if not isinstance(q_end, dt.date):
            return None
        lines = [re.sub(r"\s+", " ", ln).strip() for ln in raw_txt.splitlines() if str(ln).strip()]
        if not lines:
            return None
        in_thousands = "(in thousands)" in raw_txt.lower()
        candidates: List[Dict[str, Any]] = []
        facility_re = re.compile(
            r"\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?)\s*(million|billion)\s+"
            r"(revolver|revolving credit(?: facility)?)\b",
            re.I,
        )
        for idx, line in enumerate(lines):
            m_fac = facility_re.search(line)
            if not m_fac:
                continue
            facility = _scale_amount(m_fac.group(1), m_fac.group(2))
            if facility is None:
                continue
            tail = line[m_fac.end():]
            num_tokens = re.findall(r"\$?\s*\(?([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?)\)?", tail)
            drawn = None
            if num_tokens:
                drawn = _scale_amount(num_tokens[0], "", default_millions=False)
                if drawn is not None and in_thousands and drawn < 1_000_000:
                    drawn *= 1_000.0
            label_line = lines[idx - 1] if idx > 0 else ""
            snippet = re.sub(r"\s+", " ", f"{label_line} {line}").strip()
            facility_class, facility_priority = _classify_local_revolver_facility(
                snippet,
                doc_name=path_in.name,
                commitment=facility,
                facility_size=facility,
            )
            candidates.append(
                {
                    "quarter": q_end,
                    "facility_size": facility,
                    "commitment": facility,
                    "drawn": drawn,
                    "snippet": snippet,
                    "facility_class": facility_class,
                    "facility_priority": facility_priority,
                }
            )
        if not candidates:
            return None
        candidates.sort(
            key=lambda r: (
                r.get("facility_priority", 0),
                float(r.get("facility_size") or 0.0),
                r.get("drawn") is not None,
            ),
            reverse=True,
        )
        best = candidates[0]
        availability = None
        avail_snip = None
        avail_pats = [
            re.compile(
                r"(\$?\s*[0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?)\s*(million|billion)?\s+available\s+under\s+our\s+committed\s+revolving\s+credit\s+agreement",
                re.I,
            ),
            re.compile(
                r"(\$?\s*[0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?)\s*(million|billion)?\s+available\s+under\s+(?:our|the)\s+revolving\s+credit\s+agreement",
                re.I,
            ),
        ]
        for pat in avail_pats:
            m_av = pat.search(raw_txt)
            if not m_av:
                continue
            availability = _scale_amount(m_av.group(1).replace("$", "").strip(), m_av.group(2) or "", default_millions=True)
            if availability is not None:
                avail_snip = re.sub(r"\s+", " ", raw_txt[max(0, m_av.start() - 80):m_av.end() + 180]).strip()
                break
        letters_of_credit = None
        if (
            availability is not None
            and best.get("commitment") is not None
            and best.get("drawn") is not None
        ):
            residual = float(best["commitment"]) - float(best["drawn"]) - float(availability)
            if residual >= -1_000_000.0:
                letters_of_credit = max(residual, 0.0)
        return {
            "quarter": q_end,
            "revolver_commitment": float(best["commitment"]),
            "revolver_facility_size": float(best["facility_size"]),
            "revolver_drawn": None if best.get("drawn") is None else float(best["drawn"]),
            "revolver_letters_of_credit": None if letters_of_credit is None else float(letters_of_credit),
            "revolver_availability": None if availability is None else float(availability),
            "source_type": "local_fs_main_revolver",
            "commitment_source_type": "text",
            "facility_source_type": "text",
            "drawn_source_type": "text" if best.get("drawn") is not None else "missing",
            "lc_source_type": "derived" if letters_of_credit is not None else "missing",
            "availability_source_type": "text" if availability is not None else "missing",
            "commitment_snippet": best.get("snippet"),
            "drawn_snippet": best.get("snippet"),
            "lc_snippet": avail_snip,
            "availability_snippet": avail_snip,
            "source_snippet": avail_snip or best.get("snippet"),
            "note": f"primary revolver from local financial statement {path_in.name}",
        }

    rows: List[Dict[str, Any]] = []
    try:
        fs_files = sorted(p for p in fs_dir.iterdir() if p.is_file()) if fs_dir.exists() and fs_dir.is_dir() else []
    except Exception:
        fs_files = []
    tkr_l = str(ticker or "").strip().lower()
    for path_in in fs_files:
        if path_in.suffix.lower() not in {".pdf", ".txt", ".htm", ".html"}:
            continue
        name_l = path_in.name.lower()
        if tkr_l and tkr_l not in name_l:
            continue
        parsed = _parse_doc(path_in)
        if parsed:
            rows.append(parsed)

    narrative_rows: Dict[dt.date, Dict[str, Any]] = {}
    narrative_dirs = [
        Path(base_dir) / "earnings_transcripts",
        Path(base_dir) / "earnings_presentation",
        Path(base_dir) / "earnings_release",
    ]
    availability_patterns = [
        re.compile(
            r"\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?)\s*(million|billion|m|bn)\s+(?:in|of)\s+working capital revolver availability\b",
            re.I,
        ),
        re.compile(
            r"\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?)\s*(million|billion|m|bn)\s+available under (?:our|the)\s+(?:committed\s+)?revolving credit",
            re.I,
        ),
    ]
    for mat_dir in narrative_dirs:
        if not mat_dir.exists() or not mat_dir.is_dir():
            continue
        try:
            mat_files = sorted(p for p in mat_dir.iterdir() if p.is_file())
        except Exception:
            mat_files = []
        for path_in in mat_files:
            if path_in.suffix.lower() not in {".pdf", ".txt", ".htm", ".html"}:
                continue
            name_l = path_in.name.lower()
            if tkr_l and tkr_l not in name_l:
                continue
            raw_txt = _read_text(path_in)
            if not raw_txt:
                continue
            q_end = _parse_quarter_from_name(path_in.name) or infer_quarter_end_from_text(raw_txt)
            if not isinstance(q_end, dt.date):
                continue
            availability = None
            avail_snip = None
            for pat in availability_patterns:
                m_av = pat.search(raw_txt)
                if not m_av:
                    continue
                availability = _scale_amount(m_av.group(1), m_av.group(2) or "", default_millions=True)
                if availability is not None:
                    avail_snip = re.sub(r"\s+", " ", raw_txt[max(0, m_av.start() - 80):m_av.end() + 180]).strip()
                    break
            if availability is None:
                continue
            existing = narrative_rows.get(q_end) or {
                "quarter": q_end,
                "revolver_commitment": None,
                "revolver_facility_size": None,
                "revolver_drawn": None,
                "revolver_letters_of_credit": None,
                "revolver_availability": None,
                "source_type": "local_narrative_revolver",
                "commitment_source_type": "missing",
                "facility_source_type": "missing",
                "drawn_source_type": "missing",
                "lc_source_type": "missing",
                "availability_source_type": "missing",
                "commitment_snippet": None,
                "drawn_snippet": None,
                "lc_snippet": None,
                "availability_snippet": None,
                "source_snippet": None,
                "note": "",
            }
            existing["revolver_availability"] = float(availability)
            existing["availability_source_type"] = "text"
            existing["availability_snippet"] = avail_snip
            existing["source_snippet"] = avail_snip or existing.get("source_snippet")
            existing["note"] = f"local narrative revolver availability from {path_in.name}"
            narrative_rows[q_end] = existing

    if narrative_rows:
        if rows:
            rows_by_q: Dict[dt.date, Dict[str, Any]] = {r["quarter"]: dict(r) for r in rows if isinstance(r.get("quarter"), dt.date)}
            for q_end, narrative_row in narrative_rows.items():
                merged = rows_by_q.get(q_end, {"quarter": q_end})
                merged.setdefault("revolver_commitment", None)
                merged.setdefault("revolver_facility_size", None)
                merged.setdefault("revolver_drawn", None)
                merged.setdefault("revolver_letters_of_credit", None)
                merged.setdefault("source_type", narrative_row.get("source_type"))
                for col in (
                    "revolver_availability",
                    "availability_source_type",
                    "availability_snippet",
                    "source_snippet",
                    "note",
                ):
                    if narrative_row.get(col) not in {None, ""}:
                        merged[col] = narrative_row.get(col)
                rows_by_q[q_end] = merged
            rows = list(rows_by_q.values())
        else:
            rows = list(narrative_rows.values())

    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["quarter"] = pd.to_datetime(df["quarter"], errors="coerce")
    df = df[df["quarter"].notna()].sort_values("quarter").drop_duplicates(subset=["quarter"], keep="last")
    return df


def build_debt_buckets(
    debt_tranches_latest: pd.DataFrame,
    hist: pd.DataFrame,
    maturity_df: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    dt = debt_tranches_latest.copy() if debt_tranches_latest is not None and not debt_tranches_latest.empty else pd.DataFrame()
    mat = maturity_df.copy() if maturity_df is not None and not maturity_df.empty else pd.DataFrame()
    if not dt.empty:
        as_of = pd.to_datetime(dt["quarter"], errors="coerce").max()
    elif not mat.empty:
        as_of = pd.to_datetime(mat["quarter"], errors="coerce").max()
    else:
        return pd.DataFrame(), pd.DataFrame()
    if pd.isna(as_of):
        return pd.DataFrame(), pd.DataFrame()
    buckets = {"2026": 0.0, "2027": 0.0, "2028+": 0.0, "Unknown": 0.0}
    tranche_sum = None
    table_total_debt = None
    bucket_source = "Debt_Tranches_Latest"
    if not dt.empty:
        dt["maturity_year"] = pd.to_numeric(dt.get("maturity_year"), errors="coerce")
        amt_col = "amount_principal" if "amount_principal" in dt.columns else "amount"
        valid_amt = pd.to_numeric(dt.get(amt_col), errors="coerce")
        valid_amt = valid_amt.where(valid_amt.notna(), None)
        has_valid_tranches = bool(valid_amt.notna().any())
        if has_valid_tranches:
            for _, r in dt.iterrows():
                amt = pd.to_numeric(r.get(amt_col), errors="coerce")
                yr = r.get("maturity_year")
                if pd.isna(amt):
                    continue
                if pd.isna(yr):
                    buckets["Unknown"] += float(amt)
                elif int(yr) <= 2026:
                    buckets["2026"] += float(amt)
                elif int(yr) == 2027:
                    buckets["2027"] += float(amt)
                else:
                    buckets["2028+"] += float(amt)
            tranche_sum = float(pd.to_numeric(dt.get(amt_col), errors="coerce").sum())
            if "table_total_debt" in dt.columns:
                tt = dt.copy()
                if "period_match" in tt.columns:
                    pm = tt["period_match"].fillna(False).astype(bool)
                    if pm.any():
                        tt = tt.loc[pm].copy()
                vals = pd.to_numeric(tt.get("table_total_debt"), errors="coerce").dropna()
                if not vals.empty:
                    table_total_debt = float(vals.median())
    if tranche_sum is None and not mat.empty:
        bucket_source = str(mat.get("source_kind").astype(str).iloc[0]) if "source_kind" in mat.columns else "Debt_Maturity_Ladder"
        mat["maturity_year"] = pd.to_numeric(mat.get("maturity_year"), errors="coerce")
        for _, r in mat.iterrows():
            amt = pd.to_numeric(r.get("amount_total"), errors="coerce")
            yr = r.get("maturity_year")
            label = str(r.get("maturity_label") or "").strip().lower()
            if pd.isna(amt):
                continue
            if "thereafter" in label:
                buckets["2028+"] += float(amt)
            elif pd.isna(yr):
                buckets["Unknown"] += float(amt)
            elif int(yr) <= 2026:
                buckets["2026"] += float(amt)
            elif int(yr) == 2027:
                buckets["2027"] += float(amt)
            else:
                buckets["2028+"] += float(amt)
        tranche_sum = float(pd.to_numeric(mat.get("amount_total"), errors="coerce").dropna().sum())
    total_bucketed = sum(buckets.values())
    if table_total_debt is None and not dt.empty and "table_total_debt" in dt.columns:
        tt = dt.copy()
        if "period_match" in tt.columns:
            pm = tt["period_match"].fillna(False).astype(bool)
            if pm.any():
                tt = tt.loc[pm].copy()
        vals = pd.to_numeric(tt.get("table_total_debt"), errors="coerce").dropna()
        if not vals.empty:
            table_total_debt = float(vals.median())
    debt_core = None
    debt_long_term = None
    if hist is not None and not hist.empty and "debt_core" in hist.columns:
        h = hist.copy()
        h["quarter"] = pd.to_datetime(h["quarter"], errors="coerce")
        row = h[h["quarter"] == as_of]
        if not row.empty:
            debt_core = pd.to_numeric(row["debt_core"].iloc[0], errors="coerce")
            debt_lt = pd.to_numeric(row.get("debt_long_term"), errors="coerce").iloc[0] if "debt_long_term" in row.columns else pd.NA
            total_debt = pd.to_numeric(row.get("total_debt"), errors="coerce").iloc[0] if "total_debt" in row.columns else pd.NA
            if pd.notna(debt_lt):
                debt_long_term = float(debt_lt)
            elif pd.notna(total_debt):
                debt_long_term = float(total_debt)
    basis_value = None
    basis_label = "debt_core"
    source_basis = bucket_source
    if bucket_source == "scheduled_repayments_fallback":
        source_basis = "principal_excl_issuance_costs"
        if debt_long_term is not None and not pd.isna(debt_long_term) and float(debt_long_term) != 0:
            basis_value = float(debt_long_term)
            basis_label = "debt_long_term"
        elif table_total_debt not in (None, 0):
            basis_value = float(table_total_debt)
            basis_label = "table_total_debt"
    if basis_value in (None, 0) and debt_core is not None and pd.notna(debt_core) and float(debt_core) != 0:
        basis_value = float(debt_core)
        basis_label = "debt_core"
    coverage_pct = None
    unknown_pct = None
    if basis_value not in (None, 0):
        coverage_pct = total_bucketed / float(basis_value)
        unknown_pct = buckets["Unknown"] / float(basis_value)
    tranche_vs_table_pct = None
    if tranche_sum is not None and table_total_debt is not None and table_total_debt != 0:
        tranche_vs_table_pct = tranche_sum / float(table_total_debt)
    df = pd.DataFrame([{
        "as_of": as_of.date(),
        "2026": buckets["2026"],
        "2027": buckets["2027"],
        "2028+": buckets["2028+"],
        "Unknown": buckets["Unknown"],
        "Total_bucketed": total_bucketed,
        "Tranche_sum": tranche_sum,
        "Table_total_debt": table_total_debt,
        "Tranche_vs_table_pct": tranche_vs_table_pct,
        "Debt_core": float(debt_core) if debt_core is not None and pd.notna(debt_core) else None,
        "Debt_long_term": float(debt_long_term) if debt_long_term is not None and pd.notna(debt_long_term) else None,
        "Bucket_coverage_pct": coverage_pct,
        "Unknown_pct": unknown_pct,
        "Source": bucket_source,
        "Source_basis": source_basis,
        "Coverage_basis_metric": basis_label,
        "Coverage_basis_value": basis_value,
    }])
    qa_rows = []
    if coverage_pct is not None:
        status = "pass"
        if coverage_pct < 0.7:
            status = "fail"
        elif coverage_pct < 0.9:
            status = "warn"
        if status != "pass":
            qa_rows.append({
                "quarter": as_of.date(),
                "metric": "debt_buckets",
                "check": "debt_bucket_coverage",
                "status": status,
                "value": coverage_pct,
                "message": f"Debt bucket coverage {coverage_pct:.2%} below threshold.",
            })
        if coverage_pct > 1.05 and bucket_source != "scheduled_repayments_fallback":
            qa_rows.append({
                "quarter": as_of.date(),
                "metric": "debt_buckets",
                "check": "debt_bucket_over",
                "status": "warn",
                "value": coverage_pct,
                "message": f"Bucketed debt exceeds {basis_label} (coverage {coverage_pct:.2%}).",
            })
    if unknown_pct is not None:
        status = "pass"
        if unknown_pct > 0.3:
            status = "fail"
        elif unknown_pct > 0.1:
            status = "warn"
        if status != "pass":
            qa_rows.append({
                "quarter": as_of.date(),
                "metric": "debt_buckets",
                "check": "debt_bucket_unknown",
                "status": status,
                "value": unknown_pct,
                "message": f"Unknown maturity bucket {unknown_pct:.2%} of debt_core.",
            })
    if tranche_vs_table_pct is not None:
        diff = abs(tranche_sum - table_total_debt) / table_total_debt if table_total_debt else None
        if diff is not None:
            status = "pass"
            if diff > 0.2:
                status = "fail"
            elif diff > 0.1:
                status = "warn"
            if status != "pass":
                qa_rows.append({
                    "quarter": as_of.date(),
                    "metric": "debt_buckets",
                    "check": "debt_tranche_vs_table",
                    "status": status,
                    "value": diff,
                    "message": f"Tranche sum vs table total diff {diff:.2%}.",
                })
    return df, pd.DataFrame(qa_rows)


def compute_long_term_debt_instant(df_all: pd.DataFrame, end: dt.date, prefer_forms: List[str]) -> Optional[PickResult]:
    for tag in ["LongTermDebt", "LongTermDebtAndCapitalLeaseObligations"]:
        rec = _pick_instant_tag(df_all, end=end, tag=tag, prefer_forms=prefer_forms)
        if rec is not None:
            return PickResult(
                value=float(rec["val"]),
                source="direct",
                tag=tag,
                accn=str(rec["accn"]),
                form=str(rec["form"]),
                filed=rec["filed_d"],
                start=rec["start_d"],
                end=rec["end_d"],
                unit=str(rec["unit"]),
                duration_days=None,
                note=f"long_term_debt from {tag}",
            )
    return None


def build_tag_coverage(df_all: pd.DataFrame) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for spec in GAAP_SPECS:
        if not spec.tags:
            continue
        for tag in spec.tags:
            sub = df_all[df_all["tag"] == tag].copy()
            if sub.empty:
                rows.append({
                    "metric": spec.name,
                    "tag": tag,
                    "rows": 0,
                    "n_3m": 0,
                    "n_6m": 0,
                    "n_9m": 0,
                    "n_fy": 0,
                    "n_instant": 0,
                })
                continue
            if spec.kind == "instant":
                rows.append({
                    "metric": spec.name,
                    "tag": tag,
                    "rows": len(sub),
                    "n_3m": 0,
                    "n_6m": 0,
                    "n_9m": 0,
                    "n_fy": 0,
                    "n_instant": int(sub["end_d"].notna().sum()),
                })
            else:
                sub = sub[sub["start_d"].notna() & sub["end_d"].notna()].copy()
                # Ensure datetime-like to avoid .dt accessor errors
                sub["start_d"] = pd.to_datetime(sub["start_d"], errors="coerce")
                sub["end_d"] = pd.to_datetime(sub["end_d"], errors="coerce")
                sub = sub[sub["start_d"].notna() & sub["end_d"].notna()].copy()
                sub["dur"] = (sub["end_d"] - sub["start_d"]).dt.days
                sub["dur_class"] = sub["dur"].apply(classify_duration)
                rows.append({
                    "metric": spec.name,
                    "tag": tag,
                    "rows": len(sub),
                    "n_3m": int((sub["dur_class"] == "3M").sum()),
                    "n_6m": int((sub["dur_class"] == "6M").sum()),
                    "n_9m": int((sub["dur_class"] == "9M").sum()),
                    "n_fy": int((sub["dur_class"] == "FY").sum()),
                    "n_instant": 0,
                })
    return pd.DataFrame(rows)


def _detect_scale_from_text(text: str) -> float:
    t = (text or "").lower()
    if "in thousands" in t or "$ in thousands" in t:
        return 1000.0
    if "in millions" in t or "$ in millions" in t:
        return 1_000_000.0
    return 1.0


def _parse_header_dates_from_table(df: pd.DataFrame) -> Dict[int, pd.Timestamp]:
    def _parse_date(s: str) -> Optional[pd.Timestamp]:
        m = re.search(
            r"(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|"
            r"Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\.?\s+\d{1,2},?\s+\d{4}",
            s,
            re.IGNORECASE,
        )
        if not m:
            return None
        try:
            return pd.to_datetime(m.group(0), errors="coerce").date()
        except Exception:
            return None

    col_dates: Dict[int, pd.Timestamp] = {}
    cols = [str(c) for c in df.columns]
    for i, c in enumerate(cols):
        d = _parse_date(c)
        if d:
            col_dates[i] = d
    if col_dates:
        return col_dates

    # sometimes header dates are in the first few rows
    if not df.empty:
        for ridx in range(min(4, len(df))):
            row = [str(x) for x in df.iloc[ridx].tolist()]
            for i, c in enumerate(row):
                d = _parse_date(c)
                if d:
                    col_dates[i] = d
    return col_dates


def _source_class(src: str) -> str:
    s = (src or "").lower()
    if s in {"direct"}:
        return "xbrl_fact"
    if s.startswith("derived_") or s in {"derived_parts"}:
        return "xbrl_derived"
    if s in {"tier2_table", "tier3_10k_quarterly_data", "tier3_cash_taxes", "derived_ytd_tax_paid"}:
        return "filing_table"
    if "ex99" in s:
        if "ocr" in s:
            return "exhibit_ocr_image"
        if "pdf" in s:
            return "exhibit_pdf_text"
        return "exhibit_text"
    if s in {"missing"}:
        return "missing"
    return "other"


def _source_method(src: str) -> str:
    s = (src or "").lower()
    if s in {"direct"}:
        return "direct"
    if s.startswith("derived_"):
        return s
    if s == "tier2_table":
        return "table_strict"
    if s == "tier3_10k_quarterly_data":
        return "10k_qfd_parse"
    if s == "tier3_cash_taxes":
        return "table_supplemental"
    if s == "derived_ytd_tax_paid":
        return "derived_ytd"
    if "ex99" in s and "ocr" in s:
        return "ex99_ocr"
    if "ex99" in s:
        return "ex99_parse"
    if s == "missing":
        return "missing"
    return s or "unknown"


def _source_qa(src: str) -> str:
    cls = _source_class(src)
    if cls == "missing":
        return "FAIL"
    if str(src or "").lower() == "carry_forward":
        return "PASS"
    if cls == "exhibit_ocr_image":
        return "WARN"
    if cls == "other":
        return "WARN"
    return "PASS"


def _source_label(src: str) -> str:
    cls = _source_class(src)
    method = _source_method(src)
    if str(src or "").lower() == "carry_forward":
        return "Carry forward"
    if cls == "xbrl_fact":
        return "SEC fact"
    if cls == "xbrl_derived":
        return "SEC derived"
    if cls == "filing_table":
        if method == "10k_qfd_parse":
            return "10-K QFD"
        if method == "table_supplemental":
            return "10-Q supplemental"
        return "10-Q/10-K table"
    if cls == "exhibit_text":
        return "EX-99"
    if cls == "exhibit_pdf_text":
        return "EX-99 PDF"
    if cls == "exhibit_ocr_image":
        return "EX-99 OCR"
    if cls == "missing":
        return "Missing"
    return "Other"


def _source_tier(src: str) -> str:
    cls = _source_class(src)
    if cls in {"xbrl_fact", "xbrl_derived"}:
        return "Tier1"
    if cls == "filing_table":
        return "Tier2"
    if cls.startswith("exhibit"):
        return "Tier3"
    if cls == "missing":
        return "Missing"
    return "Other"


def _stringify_table_cells(values: Iterable[Any]) -> List[str]:
    out: List[str] = []
    for value in values:
        if pd.isna(value):
            out.append("")
        else:
            out.append(str(value))
    return out


def _table_head_text(frame: pd.DataFrame, rows: int) -> str:
    return " ".join(_stringify_table_cells(frame.head(rows).values.ravel().tolist())).lower()


def _extract_income_statement_from_html(
    html_bytes: bytes,
    quarter_end: dt.date,
    rules: Optional[Dict[str, Any]] = None,
    *,
    period_hint: str = "3M",
    parsed_bundle: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    bundle = parsed_bundle if isinstance(parsed_bundle, dict) else _parse_primary_filing_html_bundle(html_bytes)
    html = str(bundle.get("html") or "")
    scale = float(bundle.get("scale") or 1.0)
    tables = list(bundle.get("tables") or [])

    rules = rules or get_income_statement_rules(None)
    titles_any = [s.lower() for s in rules.get("titles_any", [])]
    if period_hint == "3M":
        period_markers = [s.lower() for s in rules.get("period_markers", [])]
        period_tokens = ["three months", "three-months"]
    elif period_hint == "9M":
        period_markers = ["nine months ended", "nine-months ended", "nine months", "nine-months"]
        period_tokens = ["nine months", "nine-months"]
    elif period_hint == "FY":
        period_markers = ["year ended", "years ended", "twelve months ended", "twelve-months ended", "fiscal year"]
        period_tokens = ["year ended", "years ended", "twelve months", "fiscal year"]
    else:
        period_markers = [s.lower() for s in rules.get("period_markers", [])]
        period_tokens = []
    require_labels = [s.lower() for s in rules.get("require_labels", [])]
    revenue_primary = [s.lower() for s in rules.get("revenue_primary", [])]
    revenue_parts = [s.lower() for s in rules.get("revenue_parts", [])]
    cogs_parts = [s.lower() for s in rules.get("cogs_parts", [])]
    cogs_parts_sets = [
        [s.lower() for s in part_set]
        for part_set in rules.get("cogs_parts_sets", [])
        if part_set
    ]
    cogs_alt = [s.lower() for s in rules.get("cogs_alt", [])]
    anti_labels = [s.lower() for s in rules.get("anti_labels", [])]
    cogs_min_ratio = float(rules.get("cogs_min_ratio", 0.15))
    cogs_max_ratio = float(rules.get("cogs_max_ratio", 0.95))

    def _norm_label(s: str) -> str:
        s = s.lower()
        s = re.sub(r"[^a-z0-9]+", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _match_priority(label_norm: str, priorities: List[str]) -> Optional[int]:
        for i, p in enumerate(priorities):
            if p in label_norm:
                return i
        return None

    def _col_texts(t2: pd.DataFrame) -> List[str]:
        texts: List[str] = []
        row0 = [str(x) for x in t2.columns]
        row1 = [str(x) for x in t2.iloc[0].tolist()] if len(t2) > 0 else [""] * len(row0)
        row2 = [str(x) for x in t2.iloc[1].tolist()] if len(t2) > 1 else [""] * len(row0)
        for i in range(len(row0)):
            texts.append(" ".join([row0[i], row1[i], row2[i]]).lower())
        return texts

    def _find_year_header_row(t2: pd.DataFrame, tokens: List[str]) -> Optional[Tuple[int, Dict[int, int]]]:
        max_rows = min(4, len(t2))
        for r in range(max_rows):
            row = [str(x) for x in t2.iloc[r].tolist()]
            row_l = " ".join(row).lower()
            year_cols: Dict[int, int] = {}
            token_row = bool(tokens and any(tok in row_l for tok in tokens))
            for i, cell in enumerate(row):
                m = re.search(r"\b(20\d{2})\b", str(cell))
                if m:
                    year_cols[i] = int(m.group(1))
            year_row_idx = r
            data_start = r + 1
            if token_row and len(year_cols) < 2 and r + 1 < len(t2):
                # Handle two-row headers: period text in row r, years in row r+1
                row2 = [str(x) for x in t2.iloc[r + 1].tolist()]
                year_cols = {}
                for i, cell in enumerate(row2):
                    m = re.search(r"\b(20\d{2})\b", str(cell))
                    if m:
                        year_cols[i] = int(m.group(1))
                year_row_idx = r + 1
                data_start = r + 2

            if len(year_cols) >= 2:
                # Map year labels to numeric value columns (sometimes shifted)
                year_to_col: Dict[int, int] = {}
                last_value_col = -1
                for col_idx in sorted(year_cols.keys()):
                    yr = year_cols[col_idx]
                    value_col = None
                    start_j = max(col_idx, last_value_col + 1)
                    for j in range(start_j, min(col_idx + 6, t2.shape[1])):
                        vals = [coerce_number(v) for v in t2.iloc[data_start : min(data_start + 7, len(t2)), j]]
                        if sum(v is not None for v in vals) >= 2:
                            value_col = j
                            break
                    if value_col is None:
                        value_col = col_idx
                    year_to_col[value_col] = yr
                    last_value_col = value_col
                return data_start, year_to_col
        return None

    best = None
    best_score = 0
    for t in tables:
        if t is None or t.empty:
            continue
        t2 = t.copy()
        header_text = " ".join([str(c) for c in t2.columns]).lower()
        body_text = _table_head_text(t2, 30)
        table_text = header_text + " " + body_text

        title_ok = True
        if titles_any:
            title_ok = any(k in table_text for k in titles_any)
        if period_markers and not any(k in table_text for k in period_markers):
            continue
        if period_hint != "FY" and ("year ended" in header_text or "twelve months" in header_text) and ("three months" not in header_text and "quarter" not in header_text):
            # avoid FY-only tables for quarterly fallback
            continue
        # choose label col by alpha density
        alpha_scores = {}
        for c in t2.columns:
            vals = _stringify_table_cells(t2[c].head(25).tolist())
            alpha_scores[c] = sum(1 for v in vals if re.search(r"[A-Za-z]", v))
        label_col = max(alpha_scores, key=alpha_scores.get)
        # numeric columns
        num_cols = []
        for c in t2.columns:
            vals = [coerce_number(v) for v in t2[c].head(25).tolist()]
            if sum(v is not None for v in vals) >= max(3, len(vals) // 4):
                num_cols.append(c)
        if not num_cols:
            continue

        # score table for income statement markers
        hay = " ".join(_stringify_table_cells(t2[label_col].head(30).tolist())).lower()
        score = 0
        if "revenue" in hay or "net sales" in hay:
            score += 2
        if "cost of" in hay:
            score += 1
        if "gross profit" in hay:
            score += 1
        if "operating income" in hay or "income from operations" in hay:
            score += 1
        if score < 3:
            continue

        # pick column: strict match on quarter_end if present
        col_dates = _parse_header_dates_from_table(t2)
        col_texts = _col_texts(t2)
        col_idx = None
        data_start_row = 0
        if col_dates:
            candidates = [i for i, d in col_dates.items() if d == quarter_end]
            if candidates:
                prefer = []
                if period_tokens:
                    for i in candidates:
                        if any(tok in col_texts[i] for tok in period_tokens):
                            prefer.append(i)
                if prefer:
                    col_idx = t2.columns[prefer[0]]
                elif period_hint in ("9M", "FY"):
                    col_idx = None
                else:
                    col_idx = t2.columns[candidates[0]]
        if col_idx is None and len(t2) >= 3:
            # Fallback: detect "Three Months Ended" header with year row beneath
            row1 = [str(x) for x in t2.iloc[1].tolist()]
            row2 = [str(x) for x in t2.iloc[2].tolist()]
            for i in range(min(len(row1), len(t2.columns))):
                r1 = row1[i].lower()
                if period_hint == "3M" and "three months ended" in r1 and str(quarter_end.year) in row2[i]:
                    col_idx = t2.columns[i]
                    break
                if period_hint == "9M" and ("nine months ended" in r1 or "nine-months ended" in r1) and str(quarter_end.year) in row2[i]:
                    col_idx = t2.columns[i]
                    break
                if period_hint == "FY" and ("year ended" in r1 or "twelve months ended" in r1) and str(quarter_end.year) in row2[i]:
                    col_idx = t2.columns[i]
                    break
        if col_idx is None and period_hint in ("9M", "FY"):
            hdr = _find_year_header_row(t2, period_tokens)
            if hdr is not None:
                data_start_row, year_cols = hdr
                cands = [i for i, yr in year_cols.items() if yr == quarter_end.year]
                if cands:
                    # For 9M, prefer the rightmost column (often the 9M block)
                    pick_idx = max(cands) if period_hint in ("9M", "FY") else min(cands)
                    col_idx = t2.columns[pick_idx]
        if col_idx is None:
            if period_hint in ("9M", "FY"):
                continue
            # fallback: use first numeric column after label
            col_idx = num_cols[0]

        rows = t2[[label_col, col_idx]].copy()
        if data_start_row:
            rows = rows.iloc[data_start_row:].copy()
        rows[label_col] = rows[label_col].astype(str)
        vals: Dict[str, float] = {}
        labels: Dict[str, str] = {}
        cogs_candidates: List[Tuple[str, float]] = []
        gp_candidates: List[Tuple[str, float]] = []
        op_candidates: List[Tuple[str, float]] = []
        total_costs_val = None
        total_costs_label = ""

        # build label index for required checks
        norm_labels = []
        for _, r in rows.iterrows():
            label = str(r[label_col]).strip()
            if not label or label.lower() == "nan":
                continue
            norm_labels.append(_norm_label(label))
        if require_labels:
            ok = True
            for req in require_labels:
                parts = [p.strip() for p in req.split("|")]
                if not any(any(p in lab for lab in norm_labels) for p in parts):
                    ok = False
                    break
            if not ok and period_hint in ("FY", "9M"):
                has_total_rev = any("total revenue" in lab or "total revenues" in lab or "revenues" == lab for lab in norm_labels)
                has_cogs_marker = any(("cost of" in lab or "cost of sales" in lab) for lab in norm_labels)
                ok = has_total_rev and has_cogs_marker
            if not ok:
                continue
        if not title_ok:
            has_total_rev = any("total revenue" in lab for lab in norm_labels)
            has_total_costs = any("total costs and expenses" in lab for lab in norm_labels)
            has_cogs_marker = any(("cost of" in lab or "cost of sales" in lab) for lab in norm_labels)
            # allow title-missing tables only if they are clearly IS (total revenue + total costs + cost rows)
            if not (has_total_rev and has_total_costs and has_cogs_marker):
                continue

        for _, r in rows.iterrows():
            raw_label = r[label_col]
            label = "" if pd.isna(raw_label) else str(raw_label).strip().lower()
            if not label or label == "nan":
                continue
            label_norm = _norm_label(label)
            if any(bad in label_norm for bad in anti_labels):
                # allow if this is an explicit COGS component
                if cogs_parts and any(p in label_norm for p in cogs_parts):
                    pass
                elif cogs_parts_sets and any(any(p in label_norm for p in part_set) for part_set in cogs_parts_sets):
                    pass
                else:
                    continue
            if "year ended" in label_norm or "years ended" in label_norm or "months ended" in label_norm:
                continue
            v = coerce_number(r[col_idx])
            if v is None:
                continue
            v = float(v) * scale
            # revenue primary
            if revenue_primary and _match_priority(label_norm, revenue_primary) is not None:
                vals["revenue"] = v
                labels["revenue"] = label_norm
            # track revenue parts
            if label_norm in revenue_parts:
                vals[f"rev_part::{label_norm}"] = v
                labels[f"rev_part::{label_norm}"] = label_norm
            if "total costs and expenses" in label_norm:
                total_costs_val = v
                total_costs_label = label_norm
            if "cost of" in label_norm or "cost of sales" in label_norm or "cost of revenue" in label_norm:
                cogs_candidates.append((label_norm, v))
            if "gross profit" in label_norm:
                gp_candidates.append((label, v))
            if "operating income" in label_norm or "operating loss" in label_norm or "income from operations" in label_norm or "loss from operations" in label_norm:
                op_candidates.append((label, v))

        def _pick_best_cost(cands: List[Tuple[str, float]]) -> Optional[Tuple[float, str]]:
            if not cands:
                return None
            best_v = None
            best_label = None
            best_score = None
            for label, v in cands:
                score = 0
                if re.search(r"cost of (revenue|sales|services|business services)", label):
                    score += 5
                if "total" in label:
                    score += 3
                if "cost of revenue" in label or "cost of sales" in label:
                    score += 3
                if "cost of services" in label or "cost of products" in label:
                    score += 2
                # penalize obvious sub-lines
                if any(k in label for k in ["rental", "financing", "interest", "amort", "depre", "restruct", "pension"]):
                    score -= 3
                # tie-breaker by magnitude
                mag = abs(v)
                key = (score, mag)
                if best_score is None or key > best_score:
                    best_score = key
                    best_v = v
                    best_label = label
            if best_v is None:
                return None
            return best_v, (best_label or "")

        def _pick_best_simple(cands: List[Tuple[str, float]]) -> Optional[Tuple[float, str]]:
            if not cands:
                return None
            # prefer "total" if available, else max magnitude
            totals = [(label, v) for label, v in cands if "total" in label]
            if totals:
                label, v = max(totals, key=lambda x: abs(x[1]))
                return v, label
            label, v = max(cands, key=lambda x: abs(x[1]))
            return v, label

        # Revenue resolution
        if "revenue" not in vals and revenue_parts:
            if all(f"rev_part::{p}" in vals for p in revenue_parts):
                vals["revenue"] = sum(vals[f"rev_part::{p}"] for p in revenue_parts)
                labels["revenue"] = " + ".join(revenue_parts)
        if "revenue" not in vals:
            continue

        # COGS resolution
        cogs_val = None
        cogs_label = ""
        if cogs_parts_sets:
            for part_set in cogs_parts_sets:
                if all(any(p in (l or "") for l, _ in cogs_candidates) for p in part_set):
                    parts = []
                    for p in part_set:
                        for l, v in cogs_candidates:
                            if p in l:
                                parts.append((p, v))
                                break
                    if len(parts) == len(part_set):
                        cogs_val = sum(v for _, v in parts)
                        cogs_label = " + ".join([p for p, _ in parts])
                        break
        if cogs_val is None and cogs_parts and all(any(p in (l or "") for l, _ in cogs_candidates) for p in cogs_parts):
            parts = []
            for p in cogs_parts:
                for l, v in cogs_candidates:
                    if p in l:
                        parts.append((p, v))
                        break
            if len(parts) == len(cogs_parts):
                cogs_val = sum(v for _, v in parts)
                cogs_label = " + ".join([p for p, _ in parts])
        if cogs_val is None and cogs_alt:
            alt = None
            for p in cogs_alt:
                for l, v in cogs_candidates:
                    if p in l:
                        alt = (v, l)
                        break
                if alt:
                    break
            if alt:
                cogs_val, cogs_label = alt

        if cogs_val is not None:
            vals["cogs"] = cogs_val
            labels["cogs"] = cogs_label

        # Guardrails
        if "cogs" in vals and "revenue" in vals:
            ratio = abs(vals["cogs"] / vals["revenue"]) if vals["revenue"] else 0
            if ratio < cogs_min_ratio or ratio > cogs_max_ratio:
                continue
            if total_costs_val is not None and total_costs_val < vals["cogs"]:
                continue

        cogs_pick = _pick_best_cost(cogs_candidates)
        if "cogs" not in vals and cogs_pick is not None:
            vals["cogs"], labels["cogs"] = cogs_pick
        gp_pick = _pick_best_simple(gp_candidates)
        if gp_pick is not None:
            vals["gross_profit"], labels["gross_profit"] = gp_pick
        op_pick = _pick_best_simple(op_candidates)
        if op_pick is not None:
            vals["op_income"], labels["op_income"] = op_pick

        if vals:
            if not title_ok:
                labels["_title_match"] = "title_missing"
            if score > best_score:
                best_score = score
                best = {"values": vals, "labels": labels}

    return best


def build_income_statement_fallback(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_quarters: int,
    *,
    ticker: Optional[str] = None,
    filing_inventory: Optional[Dict[str, Any]] = None,
    filing_runtime_cache: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[dt.date, Dict[str, float]], List[Dict[str, Any]]]:
    out: Dict[dt.date, Dict[str, float]] = {}
    audit_rows: List[Dict[str, Any]] = []
    seen_accn: set[str] = set()
    filing_inventory, filing_runtime_cache = _ensure_primary_filing_inventory(
        sec,
        submissions,
        filing_inventory=filing_inventory,
        filing_runtime_cache=filing_runtime_cache,
    )
    rules = get_income_statement_rules(ticker)
    for row in filing_inventory.get("rows", []) or []:
        form = row.get("form")
        if form not in ("10-Q", "10-Q/A"):
            continue
        accn = row.get("accn")
        if accn in seen_accn:
            continue
        doc = row.get("primary_doc")
        if not doc:
            continue
        q_end = row.get("report_date") or row.get("filing_date")
        if q_end is None:
            continue
        accn_nd = str(row.get("accn_nd") or "")
        seen_accn.add(str(accn))
        result = _extract_income_statement_from_primary_doc_cached(
            sec,
            cik_int,
            accn_nd,
            str(doc),
            q_end,
            filing_runtime_cache,
            rules=rules,
            period_hint="3M",
        )
        if not result:
            continue
        out[q_end] = result
        audit_rows.append({
            "quarter": q_end,
            "source": "tier2_table",
            "accn": accn,
            "doc": doc,
            "note": "income statement fallback from 10-Q/10-K table",
        })

    # trim to recent quarters
    qs = sorted(out.keys())[-max_quarters:]
    out = {q: out[q] for q in qs}
    return out, audit_rows


def build_income_statement_ytd_q4_fallback(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_quarters: int,
    *,
    ticker: Optional[str] = None,
    target_quarters: Optional[set[dt.date]] = None,
    filing_inventory: Optional[Dict[str, Any]] = None,
    filing_runtime_cache: Optional[Dict[str, Any]] = None,
    stage_timings: Optional[Dict[str, float]] = None,
    profile_timings: bool = False,
) -> Tuple[Dict[dt.date, Dict[str, float]], List[Dict[str, Any]]]:
    """
    Derive Q4 (3M) from 10-K FY table minus 10-Q 9M table.
    Only used when 3M facts are missing and strict 9M/FY columns are found.
    """
    ytd_9m: Dict[dt.date, Dict[str, Any]] = {}
    fy_map: Dict[dt.date, Dict[str, Any]] = {}
    audit_rows: List[Dict[str, Any]] = []

    target_years: Optional[set[int]] = None
    if target_quarters:
        target_years = {d.year for d in target_quarters if d is not None}

    filing_inventory, filing_runtime_cache = _ensure_primary_filing_inventory(
        sec,
        submissions,
        filing_inventory=filing_inventory,
        filing_runtime_cache=filing_runtime_cache,
        target_years=target_years,
    )
    rules = get_income_statement_rules(ticker)
    stage_timings_ref = stage_timings if stage_timings is not None else {}
    with _timed_stage(
        stage_timings_ref,
        "gaap_history.income_statement_ytd_q4_fallback.inventory_select",
        enabled=profile_timings,
    ):
        selected_rows = _select_primary_filing_rows_for_ytd_q4(
            filing_inventory,
            target_quarters,
            filing_runtime_cache=filing_runtime_cache,
        )

    with _timed_stage(
        stage_timings_ref,
        "gaap_history.income_statement_ytd_q4_fallback.parse_9m",
        enabled=profile_timings,
    ):
        for q3_end, row in (selected_rows.get("q3_rows") or {}).items():
            doc = row.get("primary_doc")
            if not doc:
                continue
            accn_nd = str(row.get("accn_nd") or "")
            result = _extract_income_statement_from_primary_doc_cached(
                sec,
                cik_int,
                accn_nd,
                str(doc),
                q3_end,
                filing_runtime_cache,
                rules=rules,
                period_hint="9M",
            )
            if not result:
                continue
            ytd_9m[q3_end] = {
                "values": result.get("values", {}),
                "labels": result.get("labels", {}),
                "accn": row.get("accn"),
                "doc": doc,
                "form": row.get("form"),
                "filed": row.get("filing_date_raw"),
            }

    with _timed_stage(
        stage_timings_ref,
        "gaap_history.income_statement_ytd_q4_fallback.parse_fy",
        enabled=profile_timings,
    ):
        for fy_end, row in (selected_rows.get("fy_rows") or {}).items():
            doc = row.get("primary_doc")
            if not doc:
                continue
            accn_nd = str(row.get("accn_nd") or "")
            result = _extract_income_statement_from_primary_doc_cached(
                sec,
                cik_int,
                accn_nd,
                str(doc),
                fy_end,
                filing_runtime_cache,
                rules=rules,
                period_hint="FY",
            )
            if not result:
                continue
            fy_map[fy_end] = {
                "values": result.get("values", {}),
                "labels": result.get("labels", {}),
                "accn": row.get("accn"),
                "doc": doc,
                "form": row.get("form"),
                "filed": row.get("filing_date_raw"),
            }

    out: Dict[dt.date, Dict[str, float]] = {}
    with _timed_stage(
        stage_timings_ref,
        "gaap_history.income_statement_ytd_q4_fallback.combine_q4",
        enabled=profile_timings,
    ):
        for fy_end, fy_payload in fy_map.items():
            q_ends = _quarter_ends_from_fy_end(fy_end)
            q3_end = q_ends.get(3)
            if not q3_end:
                continue
            ytd_payload = ytd_9m.get(q3_end)
            if not ytd_payload:
                continue
            if target_quarters is not None and fy_end not in target_quarters:
                continue
            row: Dict[str, float] = {}
            for metric in ("revenue", "cogs", "gross_profit", "op_income", "net_income"):
                fy_val = fy_payload["values"].get(metric)
                ytd_val = ytd_payload["values"].get(metric)
                if fy_val is None or ytd_val is None:
                    continue
                q4_val = float(fy_val) - float(ytd_val)
                if metric in ("revenue", "cogs") and q4_val <= 0:
                    continue
                row[metric] = q4_val
                audit_rows.append({
                    "metric": metric,
                    "quarter": fy_end,
                    "source": "derived_ytd_q4_table",
                    "tag": "10-K FY - 10-Q 9M",
                    "accn": fy_payload.get("accn"),
                    "form": fy_payload.get("form"),
                    "filed": fy_payload.get("filed"),
                    "start": None,
                    "end": fy_end,
                    "unit": "USD",
                    "duration_days": None,
                    "value": q4_val,
                    "note": f"Q4 = FY ({fy_payload.get('accn')}) - 9M ({ytd_payload.get('accn')}) from statement tables",
                })
            if row:
                out[fy_end] = row

    if out:
        qs = sorted(out.keys())[-max_quarters:]
        out = {q: out[q] for q in qs}
    return out, audit_rows


def build_quarterly_data_10k_fallback(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_quarters: int,
    *,
    target_years: Optional[set[int]] = None,
    filing_inventory: Optional[Dict[str, Any]] = None,
    filing_runtime_cache: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[dt.date, Dict[str, float]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    out: Dict[dt.date, Dict[str, float]] = {}
    audit_rows: List[Dict[str, Any]] = []
    preview_rows: List[Dict[str, Any]] = []
    seen_accn: set[str] = set()
    filing_inventory, filing_runtime_cache = _ensure_primary_filing_inventory(
        sec,
        submissions,
        filing_inventory=filing_inventory,
        filing_runtime_cache=filing_runtime_cache,
        target_years=target_years,
    )
    for row in filing_inventory.get("rows", []) or []:
        form = row.get("form")
        if form not in ("10-K", "10-K/A"):
            continue
        accn = row.get("accn")
        if accn in seen_accn:
            continue
        doc = row.get("primary_doc")
        if not doc:
            continue
        fy_end = row.get("report_date") or row.get("filing_date")
        if fy_end is None:
            continue
        if target_years is not None and fy_end.year not in target_years:
            continue
        q_ends = _quarter_ends_from_fy_end(fy_end)
        if not q_ends:
            continue
        accn_nd = str(row.get("accn_nd") or "")
        html_bytes = _load_primary_filing_document_bytes(sec, cik_int, accn_nd, str(doc), filing_runtime_cache)
        if html_bytes is None:
            continue
        seen_accn.add(str(accn))
        result = _extract_quarterly_data_10k(html_bytes)
        if not result:
            continue
        values = result.get("values", {})
        for q_idx, q_end in q_ends.items():
            row_out = out.get(q_end, {})
            for metric, qmap in values.items():
                if q_idx in qmap:
                    row_out[metric] = float(qmap[q_idx])
                    preview_rows.append({
                        "accn": accn,
                        "doc": doc,
                        "fy_end": fy_end,
                        "quarter": q_end,
                        "metric": metric,
                        "value": float(qmap[q_idx]),
                        "source": "10-K Quarterly Financial Data",
                    })
            out[q_end] = row_out
        audit_rows.append({
            "quarter": fy_end,
            "source": "tier3_10k_quarterly_data",
            "accn": accn,
            "doc": doc,
            "note": "10-K Quarterly Financial Data table",
        })

    if out:
        qs = sorted(out.keys())[-max_quarters:]
        out = {q: out[q] for q in qs}
    return out, audit_rows, preview_rows


def _detect_scale_text(txt: str) -> float:
    t = (txt or "").lower()
    if re.search(r"in\s+thousands", t):
        return 1000.0
    if re.search(r"in\s+millions", t):
        return 1_000_000.0
    return 1.0


def _first_number_from_line(line: str) -> Optional[float]:
    m = re.search(r"(-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d{4,}(?:\.\d+)?)", line)
    if not m:
        return None
    try:
        return float(m.group(0).replace(",", ""))
    except Exception:
        return None


def _extract_income_statement_from_text(
    text: str,
    quarter_end: dt.date,
    *,
    rules: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    if not text:
        return None
    low = text.lower()
    if not any(t in low for t in rules.get("period_markers", [])):
        return None
    q_end = infer_quarter_end_from_text(text)
    if q_end is not None and q_end != quarter_end:
        return None

    scale = _detect_scale_text(text)
    lines = [re.sub(r"\s+", " ", ln).strip() for ln in text.splitlines() if ln.strip()]

    def find_value(labels: List[str]) -> Tuple[Optional[float], Optional[str]]:
        for i, ln in enumerate(lines):
            if any(lbl in ln.lower() for lbl in labels):
                v = _first_number_from_line(ln)
                if v is None and i + 1 < len(lines):
                    v = _first_number_from_line(lines[i + 1])
                if v is not None:
                    return v * scale, ln.strip()
        return None, None

    values: Dict[str, float] = {}
    labels: Dict[str, str] = {}
    tokens: Dict[str, str] = {}

    def _safe_token_float(token_in: Any) -> Optional[float]:
        tok = str(token_in or "").strip().replace(",", "")
        if not tok:
            return None
        try:
            return float(tok)
        except Exception:
            return None

    def _safe_token_float(token_in: Any) -> Optional[float]:
        tok = str(token_in or "").strip().replace(",", "")
        if not tok:
            return None
        try:
            return float(tok)
        except Exception:
            return None

    rev, rev_lbl = find_value(rules.get("revenue_primary", []))
    if rev is None:
        parts = []
        for lbl in rules.get("revenue_parts", []):
            v, l = find_value([lbl])
            if v is None:
                parts = []
                break
            parts.append(v)
            labels[f"revenue_part_{lbl}"] = l or lbl
        if parts:
            rev = float(sum(parts))
    if rev is not None:
        values["revenue"] = float(rev)
        if rev_lbl:
            labels["revenue"] = rev_lbl
            tokens["revenue"] = rev_lbl

    cogs_vals = []
    for lbl in rules.get("cogs_parts", []):
        v, l = find_value([lbl])
        if v is None:
            cogs_vals = []
            break
        cogs_vals.append(v)
        labels[f"cogs_part_{lbl}"] = l or lbl
    if cogs_vals:
        cogs = float(sum(cogs_vals))
    else:
        cogs, cogs_lbl = find_value(rules.get("cogs_alt", []))
        if cogs_lbl:
            labels["cogs"] = cogs_lbl
            tokens["cogs"] = cogs_lbl

    if cogs is not None:
        values["cogs"] = float(cogs)

    gross, gp_lbl = find_value(["gross profit"])
    if gross is None and ("revenue" in values and "cogs" in values):
        gross = float(values["revenue"] - values["cogs"])
    if gross is not None:
        values["gross_profit"] = float(gross)
        if gp_lbl:
            labels["gross_profit"] = gp_lbl
            tokens["gross_profit"] = gp_lbl

    op, op_lbl = find_value(["operating income", "income from operations"])
    if op is not None:
        values["op_income"] = float(op)
        if op_lbl:
            labels["op_income"] = op_lbl
            tokens["op_income"] = op_lbl

    ni, ni_lbl = find_value(["net income", "net loss"])
    if ni is not None:
        values["net_income"] = float(ni)
        if ni_lbl:
            labels["net_income"] = ni_lbl
            tokens["net_income"] = ni_lbl

    # Guardrails
    if "revenue" not in values:
        return None
    if "cogs" in values:
        ratio = abs(values["cogs"]) / max(abs(values["revenue"]), 1.0)
        if ratio < rules.get("cogs_min_ratio", 0.05) or ratio > rules.get("cogs_max_ratio", 0.98):
            return None

    return {"values": values, "labels": labels, "tokens": tokens, "scale": scale, "source": "tier3_ex99_ocr"}


def _quarter_ends_from_fy_end(fy_end: dt.date) -> Dict[int, dt.date]:
    try:
        q4 = pd.Timestamp(fy_end)
        q3 = (q4 - pd.DateOffset(months=3)) + pd.offsets.MonthEnd(0)
        q2 = (q4 - pd.DateOffset(months=6)) + pd.offsets.MonthEnd(0)
        q1 = (q4 - pd.DateOffset(months=9)) + pd.offsets.MonthEnd(0)
        return {1: q1.date(), 2: q2.date(), 3: q3.date(), 4: q4.date()}
    except Exception:
        return {}


def _extract_quarterly_data_10k(html_bytes: bytes) -> Optional[Dict[str, Any]]:
    html = html_bytes.decode("utf-8", errors="ignore")
    scale = _detect_scale_from_text(html)
    tables = read_html_tables_any(html_bytes)

    def _eval_tables(tables_in: List[pd.DataFrame], *, require_marker: bool) -> Optional[Dict[str, Any]]:
        best_local: Optional[Dict[str, Any]] = None
        best_score_local = 0
        for t in tables_in:
            if t is None or t.empty:
                continue
            t2 = t.copy()
            header_text = " ".join([str(c) for c in t2.columns]).lower()
            body_text = _table_head_text(t2, 30)
            table_text = header_text + " " + body_text
            marker_ok = True
            if require_marker:
                marker_ok = ("quarterly financial data" in table_text) or ("quarterly" in table_text and "unaudited" in table_text)
                if not marker_ok:
                    continue

            # detect quarter columns
            colmap: Dict[Any, int] = {}
            total_col = None
            for c in t2.columns:
                q = _quarter_from_label(str(c))
                if q:
                    colmap[c] = q
                if "total" in str(c).lower():
                    total_col = c
            # if columns don't have quarter labels, check first row
            if len(colmap) < 3 and not t2.empty:
                row0 = [str(x) for x in t2.iloc[0].tolist()]
                for c, v in zip(t2.columns, row0):
                    q = _quarter_from_label(v)
                    if q:
                        colmap[c] = q
                    if "total" in v.lower():
                        total_col = c
                if len(colmap) >= 3:
                    t2 = t2.iloc[1:].copy()
            if len(colmap) < 3:
                continue

            # choose label column by alpha density
            alpha_scores = {}
            for c in t2.columns:
                vals = _stringify_table_cells(t2[c].head(25).tolist())
                alpha_scores[c] = sum(1 for v in vals if re.search(r"[A-Za-z]", v))
            label_col = max(alpha_scores, key=alpha_scores.get)

            label_col_in_colmap = label_col in colmap
            if label_col_in_colmap:
                colmap = {k: v for k, v in colmap.items() if k != label_col}
            use_seq = False
            rows = t2[[label_col] + list(colmap.keys()) + ([total_col] if total_col else [])].copy()
            rows[label_col] = rows[label_col].astype(str)

            # If quarter header columns don't contain numeric data, fall back to sequential numeric parsing
            try:
                sample = rows.head(6)
                num_in_colmap = 0
                for c in colmap.keys():
                    series = sample[c]
                    if any(coerce_number(v) is not None for v in series.tolist()):
                        num_in_colmap += 1
                if num_in_colmap == 0:
                    use_seq = True
            except Exception:
                use_seq = False
            if label_col_in_colmap:
                use_seq = True
            if use_seq:
                rows = t2.copy()
                rows[label_col] = rows[label_col].astype(str)

            # collect metrics per quarter
            data: Dict[str, Dict[int, float]] = {}
            totals: Dict[str, float] = {}
            for _, r in rows.iterrows():
                label = r[label_col]
                if isinstance(label, pd.Series):
                    label = " ".join([str(x) for x in label.tolist() if str(x) and str(x).lower() != "nan"])
                label_str = str(label).strip()
                if not label_str or label_str.lower() == "nan":
                    continue
                lab = _norm(label_str)
                is_cost_row = "cost of" in lab or "cost" in lab
                metric = None
                for k, labs in metric_labels.items():
                    if k == "revenue" and is_cost_row:
                        continue
                    if any(l in lab for l in labs):
                        metric = k
                        break
                if not metric:
                    continue
                vals_q: Dict[int, float] = {}
                if use_seq:
                    nums = []
                    for v in r.tolist():
                        num = coerce_number(v)
                        if num is not None:
                            nums.append(float(num))
                    if len(nums) >= 4:
                        vals_q = {1: nums[0] * scale, 2: nums[1] * scale, 3: nums[2] * scale, 4: nums[3] * scale}
                        if total_col is not None and len(nums) >= 5:
                            totals[metric] = nums[4] * scale
                else:
                    for c, q in colmap.items():
                        v = coerce_number(r.get(c))
                        if v is None:
                            continue
                        vals_q[q] = float(v) * scale
                if vals_q:
                    data[metric] = vals_q
                if total_col is not None and not use_seq:
                    vtot = coerce_number(r.get(total_col))
                    if vtot is not None:
                        totals[metric] = float(vtot) * scale

            if "revenue" not in data:
                continue

            # guardrails: Q1..Q4 sum vs total if total exists
            ok = True
            for metric in ("revenue", "cogs"):
                if metric in data and metric in totals:
                    qsum = sum(data[metric].get(q, 0.0) for q in (1, 2, 3, 4))
                    tot = totals[metric]
                    if tot and abs(qsum - tot) / abs(tot) > 0.05:
                        ok = False
                        break
            if not ok:
                continue

            score = len(data)
            if score > best_score_local:
                best_score_local = score
                best_local = {"values": data, "scale": scale, "totals": totals}

        return best_local

    def _norm(s: str) -> str:
        s = s.lower()
        s = re.sub(r"[^a-z0-9]+", " ", s)
        return re.sub(r"\s+", " ", s).strip()

    def _quarter_from_label(s: str) -> Optional[int]:
        s = s.lower()
        if "first" in s or "q1" in s:
            return 1
        if "second" in s or "q2" in s:
            return 2
        if "third" in s or "q3" in s:
            return 3
        if "fourth" in s or "q4" in s:
            return 4
        return None

    metric_labels = {
        "revenue": [
            "total revenue", "total revenues", "revenues", "revenue", "net revenue",
        ],
        "cogs": [
            "cost of revenue", "cost of revenues", "cost of sales",
            "cost of services", "cost of business services",
        ],
        "op_income": [
            "operating income", "operating loss", "income from operations", "loss from operations",
        ],
        "pretax_income": [
            "income from continuing operations before income taxes",
            "income loss from continuing operations before income taxes",
            "income before income taxes", "loss before income taxes",
        ],
        "net_income": [
            "net income", "net loss",
        ],
        "diluted_eps": [
            "diluted earnings per share", "diluted eps",
        ],
    }

    best = _eval_tables(tables, require_marker=True)
    if best:
        return best

    # If marker isn't inside table, find nearest table after the "Quarterly Financial Data" text
    m = re.search(r"quarterly financial data.{0,200000}?<table.*?</table>", html, flags=re.I | re.S)
    if m:
        snippet = m.group(0)
        snip_tables = read_html_tables_any(snippet.encode("utf-8", errors="ignore"))
        best = _eval_tables(snip_tables, require_marker=False)
        if best:
            return best

    # final fallback: try all tables without marker requirement (still guarded by Q1..Q4+total)
    return _eval_tables(tables, require_marker=False)


def _extract_balance_sheet_from_text(
    text: str,
    quarter_end: dt.date,
) -> Optional[Dict[str, Any]]:
    if not text:
        return None
    low = text.lower()
    if "balance sheet" not in low and "financial position" not in low:
        return None
    q_end = infer_quarter_end_from_text(text)
    if q_end is not None and q_end != quarter_end and abs((q_end - quarter_end).days) <= 120:
        return None
    scale = _detect_scale_text(text)
    lines = [re.sub(r"\s+", " ", ln).strip() for ln in text.splitlines() if ln.strip()]

    def _line_numeric_hits(line_in: Any) -> List[Tuple[float, int, int, bool]]:
        txt_local = re.sub(r"\(\d+\)", "", str(line_in or ""))
        pat = re.compile(
            r"(?<!\d)"
            r"(?:" 
            r"\$?\s*\(\s*[0-9]{1,4}(?:,[0-9]{3})*(?:\.\d+)?\s*\)(?:\s*(?:million|m))?"
            r"|[-+]?\$?\s*[0-9]{1,4}(?:,[0-9]{3})*(?:\.\d+)?(?:\s*(?:million|m))?"
            r")"
            r"(?!\d)",
            re.I,
        )
        out_vals: List[Tuple[float, int, int, bool]] = []
        for m in pat.finditer(txt_local):
            token = str(m.group(0) or "").strip()
            raw_digits = re.sub(r"[^\d]", "", token)
            if raw_digits.isdigit() and len(raw_digits) == 4:
                try:
                    year_val = int(raw_digits)
                except Exception:
                    year_val = 0
                if 1900 <= year_val <= 2100 and not any(ch in token for ch in "$(),.-"):
                    continue
            neg = bool(re.match(r"^\s*-\s*", token) or re.match(r"^\s*\$?\s*\(", token) or re.search(r"\(\s*\$?\s*\d", token))
            token_num = re.sub(r"(?i)\b(million|thousand|m|mm)\b", "", token)
            num_match = re.search(r"[-+]?\d+(?:,\d{3})*(?:\.\d+)?", token_num)
            if not num_match:
                continue
            try:
                val = float(num_match.group(0).replace(",", ""))
            except Exception:
                continue
            if neg and val > 0:
                val *= -1.0
            explicit_scale = None
            if re.search(r"\bbillion\b|\bbn\b", token, re.I):
                explicit_scale = 1_000_000_000.0
            elif re.search(r"\bmillion\b|\bmm\b|\bm\b", token, re.I):
                explicit_scale = 1_000_000.0
            elif re.search(r"\bthousand\b|\bk\b", token, re.I):
                explicit_scale = 1_000.0
            scaled_val = float(val) * float(explicit_scale or 1.0)
            out_vals.append((scaled_val, int(m.start()), int(m.end()), explicit_scale is not None))
        return out_vals

    def _pick_nearest_label_value(line_in: Any, labels_in: List[str]) -> Tuple[Optional[float], bool]:
        line_txt = str(line_in or "")
        low_line = line_txt.lower()
        label_hits = [(low_line.find(lbl), len(lbl)) for lbl in labels_in if lbl in low_line]
        if not label_hits:
            return None, False
        label_hits = [x for x in label_hits if x[0] >= 0]
        if not label_hits:
            return None, False
        label_pos, label_len = min(label_hits, key=lambda x: x[0])
        hits = _line_numeric_hits(line_txt)
        if not hits:
            return None, False
        if "$" not in line_txt and len(hits) < 2 and not re.search(r"\b(million|thousand|m)\b", line_txt, re.I):
            return None, False
        def _score(hit: Tuple[float, int, int, bool]) -> Tuple[int, int]:
            _val, start_idx, end_idx, _has_explicit_scale = hit
            if start_idx >= label_pos:
                gap = start_idx - (label_pos + label_len)
            else:
                gap = label_pos - end_idx
            return (abs(gap), 0 if start_idx >= label_pos else 1)
        best_hit = min(hits, key=_score)
        return float(best_hit[0]), bool(best_hit[3])

    def find_value(labels: List[str]) -> Tuple[Optional[float], Optional[str]]:
        for i, ln in enumerate(lines):
            if any(lbl in ln.lower() for lbl in labels):
                v, has_explicit_scale = _pick_nearest_label_value(ln, labels)
                if v is None and i + 1 < len(lines):
                    nxt = lines[i + 1]
                    if "$" in nxt or re.match(r"^\s*\$?\s*\(?-?\d", nxt):
                        nxt_hits = _line_numeric_hits(nxt)
                        if nxt_hits:
                            v = float(nxt_hits[0][0])
                            has_explicit_scale = bool(nxt_hits[0][3])
                if v is not None:
                    return (float(v) if has_explicit_scale else float(v) * scale), ln.strip()
        return None, None

    def find_matches(labels: List[str]) -> List[Tuple[float, str]]:
        out_hits: List[Tuple[float, str]] = []
        for i, ln in enumerate(lines):
            low_ln = ln.lower()
            if not any(lbl in low_ln for lbl in labels):
                continue
            v, has_explicit_scale = _pick_nearest_label_value(ln, labels)
            if v is None and i + 1 < len(lines):
                nxt = lines[i + 1]
                if "$" in nxt or re.match(r"^\s*\$?\s*\(?-?\d", nxt):
                    nxt_hits = _line_numeric_hits(nxt)
                    if nxt_hits:
                        v = float(nxt_hits[0][0])
                        has_explicit_scale = bool(nxt_hits[0][3])
            if v is None:
                continue
            out_hits.append(((float(v) if has_explicit_scale else float(v) * scale), ln.strip()))
        return out_hits

    def find_strict_matches(labels: List[str]) -> List[Tuple[float, str]]:
        out_hits: List[Tuple[float, str]] = []
        label_list = [str(lbl or "").strip().lower() for lbl in labels if str(lbl or "").strip()]
        for i, ln in enumerate(lines):
            low_ln = ln.lower()
            if not any(low_ln.startswith(lbl) for lbl in label_list):
                continue
            v, has_explicit_scale = _pick_nearest_label_value(ln, label_list)
            if v is None and i + 1 < len(lines):
                nxt = lines[i + 1]
                if "$" in nxt or re.match(r"^\s*\$?\s*\(?-?\d", nxt):
                    nxt_hits = _line_numeric_hits(nxt)
                    if nxt_hits:
                        v = float(nxt_hits[0][0])
                        has_explicit_scale = bool(nxt_hits[0][3])
            if v is None:
                continue
            out_hits.append(((float(v) if has_explicit_scale else float(v) * scale), ln.strip()))
        return out_hits

    def find_strict_value(labels: List[str]) -> Tuple[Optional[float], Optional[str]]:
        strict_hits = find_strict_matches(labels)
        if strict_hits:
            return strict_hits[0]
        return None, None

    values: Dict[str, float] = {}
    labels: Dict[str, str] = {}
    tokens: Dict[str, str] = {}

    def _safe_token_float(token_in: Any) -> Optional[float]:
        tok = str(token_in or "").strip().replace(",", "")
        if not tok:
            return None
        try:
            return float(tok)
        except Exception:
            return None

    cash_labels = ["cash and cash equivalents", "cash, cash equivalents"]
    cash, cash_lbl = find_strict_value(cash_labels)
    if cash is None:
        cash, cash_lbl = find_value(cash_labels)
    if cash is not None:
        values["cash"] = float(cash)
        if cash_lbl:
            labels["cash"] = cash_lbl
            tokens["cash"] = cash_lbl

    restricted_cash, restricted_cash_lbl = find_strict_value(["restricted cash"])
    if restricted_cash is None:
        restricted_cash, restricted_cash_lbl = find_value(["restricted cash"])
    if restricted_cash is not None:
        values["restricted_cash"] = float(restricted_cash)
        if restricted_cash_lbl:
            labels["restricted_cash"] = restricted_cash_lbl
            tokens["restricted_cash"] = restricted_cash_lbl

    prepaid_other, prepaid_other_lbl = find_value(
        [
            "prepaid expenses and other current assets",
            "prepaid expenses and other current",
            "prepaid expenses and other",
            "other current assets",
        ]
    )
    if prepaid_other is not None:
        values["prepaid_other_current_assets"] = float(prepaid_other)
        if prepaid_other_lbl:
            labels["prepaid_other_current_assets"] = prepaid_other_lbl
            tokens["prepaid_other_current_assets"] = prepaid_other_lbl

    deriv_hits = [
        hit
        for hit in find_strict_matches(["derivative financial instruments"])
        if "associated with" not in hit[1].lower() and " - " not in hit[1]
    ]
    if len(deriv_hits) < 2:
        deriv_hits = find_matches(["derivative financial instruments", "derivative assets", "derivative liabilities"])
    if deriv_hits:
        deriv_asset = next((hit for hit in deriv_hits if "liabil" not in hit[1].lower()), deriv_hits[0])
        if deriv_asset:
            values["derivative_assets"] = float(deriv_asset[0])
            labels["derivative_assets"] = deriv_asset[1]
            tokens["derivative_assets"] = deriv_asset[1]
        deriv_liab = next((hit for hit in deriv_hits if "liabil" in hit[1].lower()), deriv_hits[1] if len(deriv_hits) > 1 else None)
        if deriv_liab is not None:
            values["derivative_liabilities"] = float(deriv_liab[0])
            labels["derivative_liabilities"] = deriv_liab[1]
            tokens["derivative_liabilities"] = deriv_liab[1]

    rou_asset, rou_asset_lbl = find_strict_value(["operating lease right-of-use assets", "operating lease right of use assets"])
    if rou_asset is None:
        rou_asset, rou_asset_lbl = find_value(["operating lease right-of-use assets", "operating lease right of use assets"])
    if rou_asset is not None:
        values["operating_lease_rou_assets"] = float(rou_asset)
        if rou_asset_lbl:
            labels["operating_lease_rou_assets"] = rou_asset_lbl
            tokens["operating_lease_rou_assets"] = rou_asset_lbl

    dtx, dtx_lbl = find_strict_value(["deferred income taxes, net", "deferred tax assets net", "deferred income tax assets net"])
    if dtx is None:
        dtx, dtx_lbl = find_value(["deferred income taxes, net", "deferred tax assets net", "deferred income tax assets net"])
    if dtx is not None:
        values["deferred_income_taxes_net"] = float(dtx)
        if dtx_lbl:
            labels["deferred_income_taxes_net"] = dtx_lbl
            tokens["deferred_income_taxes_net"] = dtx_lbl

    total_debt, td_lbl = find_value(["total debt", "total borrowings", "total long-term debt"])
    if total_debt is not None:
        values["total_debt"] = float(total_debt)
        if td_lbl:
            labels["total_debt"] = td_lbl
            tokens["total_debt"] = td_lbl

    equity, eq_lbl = find_value([
        "total stockholders' deficit",
        "total stockholders' equity",
        "total shareholders' deficit",
        "total shareholders' equity",
        "total equity",
        "total equity (deficit)",
    ])
    if equity is not None:
        values["total_equity"] = float(equity)
        if eq_lbl:
            labels["total_equity"] = eq_lbl
            tokens["total_equity"] = eq_lbl

    goodwill = None
    gw_lbl = None
    gw_note = re.search(
        r"carrying amount of goodwill(?: attributable)?[^.]{0,320}?was\s+\$?\s*([0-9]{1,3}(?:\.\d+)?)\s+million",
        text,
        re.I | re.S,
    )
    if gw_note:
        gw_val = _safe_token_float(gw_note.group(1))
        if gw_val is not None:
            goodwill = float(gw_val) * 1_000_000.0
            gw_lbl = "carrying amount of goodwill"
    if goodwill is None:
        for ln in lines:
            low_ln = ln.lower()
            if not low_ln.startswith("goodwill"):
                continue
            m_gw_line = re.search(
                r"^goodwill[^0-9\$]{0,12}\$?\s*\(?([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?)",
                ln,
                re.I,
            )
            if not m_gw_line:
                continue
            gw_val = _safe_token_float(m_gw_line.group(1))
            if gw_val is None:
                continue
            goodwill = float(gw_val) * scale
            gw_lbl = ln.strip()
            break
    if goodwill is not None:
        values["goodwill"] = float(goodwill)
        if gw_lbl:
            labels["goodwill"] = gw_lbl
            tokens["goodwill"] = gw_note.group(0) if gw_note else gw_lbl

    intang, int_lbl = find_strict_value(["total intangible assets, net", "intangible assets, net"])
    if intang is None:
        intang, int_lbl = find_value(["total intangible assets, net", "intangible assets, net"])
    if intang is not None:
        values["intangibles"] = float(intang)
        if int_lbl:
            labels["intangibles"] = int_lbl
            tokens["intangibles"] = int_lbl
    else:
        int_note = re.search(
            r"total intangible assets,\s*net\s+\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d+)?)",
            text,
            re.I,
        )
        if int_note:
            int_val = _safe_token_float(int_note.group(1))
            if int_val is not None:
                values["intangibles"] = float(int_val) * scale
                labels["intangibles"] = "total intangible assets, net"
                tokens["intangibles"] = int_note.group(0)

    # Bank deposits (current/noncurrent)
    dep_cur, dep_lbl = find_value(["customer deposits at pitney bowes bank", "customer deposits at bank"])
    dep_non, depn_lbl = find_value(["noncurrent customer deposits at pitney bowes bank", "noncurrent customer deposits"])
    if dep_cur is not None or dep_non is not None:
        if dep_cur is not None and dep_non is not None:
            values["bank_deposits"] = float(dep_cur + dep_non)
            labels["bank_deposits"] = "customer deposits (current + noncurrent)"
            tokens["bank_deposits"] = (dep_lbl or "") + " | " + (depn_lbl or "")
        else:
            values["bank_deposits"] = float(dep_cur if dep_cur is not None else dep_non)
            labels["bank_deposits"] = "customer deposits (partial)"
            tokens["bank_deposits"] = dep_lbl or depn_lbl or ""

    # Bank finance receivables
    fin_cur, finc_lbl = find_value(["short-term finance receivables, net", "short term finance receivables, net"])
    fin_non, finn_lbl = find_value(["long-term finance receivables, net", "long term finance receivables, net"])
    fin_tot, fint_lbl = find_value(["finance receivables, net"])
    if fin_cur is not None or fin_non is not None:
        if fin_cur is not None and fin_non is not None:
            values["bank_finance_receivables"] = float(fin_cur + fin_non)
            labels["bank_finance_receivables"] = "finance receivables (short + long)"
            tokens["bank_finance_receivables"] = (finc_lbl or "") + " | " + (finn_lbl or "")
        else:
            values["bank_finance_receivables"] = float(fin_cur if fin_cur is not None else fin_non)
            labels["bank_finance_receivables"] = "finance receivables (partial)"
            tokens["bank_finance_receivables"] = finc_lbl or finn_lbl or ""
    elif fin_tot is not None:
        values["bank_finance_receivables"] = float(fin_tot)
        labels["bank_finance_receivables"] = "finance receivables (total)"
        tokens["bank_finance_receivables"] = fint_lbl or ""

    # Lease liabilities (operating + finance)
    lease_cur, lcur_lbl = find_strict_value(["current operating lease liabilities", "operating lease current liabilities"])
    if lease_cur is None:
        lease_cur, lcur_lbl = find_value(["current operating lease liabilities", "operating lease current liabilities"])
    lease_non, lnon_lbl = find_strict_value(
        [
            "noncurrent operating lease liabilities",
            "non current operating lease liabilities",
            "operating lease long-term liabilities",
            "operating lease long term liabilities",
        ]
    )
    if lease_non is None:
        lease_non, lnon_lbl = find_value(
            [
                "noncurrent operating lease liabilities",
                "non current operating lease liabilities",
                "operating lease long-term liabilities",
                "operating lease long term liabilities",
            ]
        )
    lease_fin_cur, lfcur_lbl = find_value(["current finance lease liabilities"])
    lease_fin_non, lfnon_lbl = find_value(["noncurrent finance lease liabilities", "non current finance lease liabilities"])
    lease_parts = [v for v in [lease_cur, lease_non, lease_fin_cur, lease_fin_non] if v is not None]
    if lease_parts:
        values["lease_liabilities"] = float(sum(lease_parts))
        labels["lease_liabilities"] = "lease liabilities (operating + finance)"
        tokens["lease_liabilities"] = " | ".join([x for x in [lcur_lbl, lnon_lbl, lfcur_lbl, lfnon_lbl] if x])
    if lease_cur is not None:
        values["operating_lease_current_liabilities"] = float(lease_cur)
        if lcur_lbl:
            labels["operating_lease_current_liabilities"] = lcur_lbl
            tokens["operating_lease_current_liabilities"] = lcur_lbl
    if lease_non is not None:
        values["operating_lease_long_term_liabilities"] = float(lease_non)
        if lnon_lbl:
            labels["operating_lease_long_term_liabilities"] = lnon_lbl
            tokens["operating_lease_long_term_liabilities"] = lnon_lbl

    st_notes, st_notes_lbl = find_strict_value(["short-term notes payable and other borrowings", "short term notes payable and other borrowings"])
    if st_notes is None:
        st_notes, st_notes_lbl = find_value(["short-term notes payable and other borrowings", "short term notes payable and other borrowings"])
    if st_notes is not None:
        values["short_term_notes_payable_and_other_borrowings"] = float(st_notes)
        if st_notes_lbl:
            labels["short_term_notes_payable_and_other_borrowings"] = st_notes_lbl
            tokens["short_term_notes_payable_and_other_borrowings"] = st_notes_lbl

    # Debt core components
    debt_cur, dcur_lbl = find_strict_value(
        [
            "current portion of long-term debt",
            "current portion of long term debt",
            "current maturities of long-term debt",
            "current maturities of long term debt",
        ]
    )
    if debt_cur is None:
        debt_cur, dcur_lbl = find_value(
            [
                "current portion of long-term debt",
                "current portion of long term debt",
                "current maturities of long-term debt",
                "current maturities of long term debt",
            ]
        )
    debt_non, dnon_lbl = find_strict_value(
        [
            "long-term debt, net of current portion",
            "long term debt, net of current portion",
            "long-term debt, less current maturities",
            "long term debt, less current maturities",
            "long-term debt noncurrent",
            "long term debt noncurrent",
        ]
    )
    if debt_non is None:
        debt_non, dnon_lbl = find_value(
            [
                "long-term debt, net of current portion",
                "long term debt, net of current portion",
                "long-term debt, less current maturities",
                "long term debt, less current maturities",
                "long-term debt noncurrent",
                "long term debt noncurrent",
            ]
        )
    debt_lt, dlt_lbl = find_strict_value(["long-term debt", "long term debt"])
    if debt_lt is None:
        debt_lt, dlt_lbl = find_value(["long-term debt", "long term debt"])
    if debt_cur is not None:
        values["current_maturities_of_long_term_debt"] = float(debt_cur)
        if dcur_lbl:
            labels["current_maturities_of_long_term_debt"] = dcur_lbl
            tokens["current_maturities_of_long_term_debt"] = dcur_lbl
    debt_val, debt_choice = choose_total_debt(
        noncurrent=float(debt_non) if debt_non is not None else None,
        current=float(debt_cur) if debt_cur is not None else None,
        longtermdebt=float(debt_lt) if debt_lt is not None else None,
    )
    if debt_val is not None:
        values["debt_core"] = float(debt_val)
        if debt_choice == "noncurrent_plus_current":
            labels["debt_core"] = "noncurrent long-term debt + current portion"
            tokens["debt_core"] = " | ".join([x for x in [dnon_lbl, dcur_lbl] if x])
        elif debt_choice == "longterm_includes_current_or_total":
            labels["debt_core"] = "long-term debt (single line)"
            tokens["debt_core"] = dlt_lbl or ""
        elif debt_choice == "noncurrent_only":
            labels["debt_core"] = "noncurrent long-term debt"
            tokens["debt_core"] = dnon_lbl or ""
        else:
            labels["debt_core"] = "current debt (fallback)"
            tokens["debt_core"] = dcur_lbl or ""

    carbon_liab, carbon_liab_lbl = find_strict_value(["carbon equipment liabilities"])
    if carbon_liab is None:
        carbon_liab, carbon_liab_lbl = find_value(["carbon equipment liabilities"])
    if carbon_liab is not None:
        values["carbon_equipment_liabilities"] = float(carbon_liab)
        if carbon_liab_lbl:
            labels["carbon_equipment_liabilities"] = carbon_liab_lbl
            tokens["carbon_equipment_liabilities"] = carbon_liab_lbl

    other_liab, other_liab_lbl = find_strict_value(["other liabilities", "other long-term liabilities", "other noncurrent liabilities"])
    if other_liab is None:
        other_liab, other_liab_lbl = find_value(["other long-term liabilities", "other noncurrent liabilities"])
    if other_liab is not None:
        values["other_liabilities"] = float(other_liab)
        if other_liab_lbl:
            labels["other_liabilities"] = other_liab_lbl
            tokens["other_liabilities"] = other_liab_lbl

    # Guardrails: if we only captured one value, that's fine. Reject obvious nonsense.
    for k in list(values.keys()):
        v = values[k]
        if v is None or not (-1e12 < v < 1e12):
            values.pop(k, None)
            tokens.pop(k, None)
            labels.pop(k, None)

    if not values:
        return None

    return {"values": values, "labels": labels, "tokens": tokens, "scale": scale, "source": "tier3_ex99_ocr"}


def _extract_balance_sheet_from_html(
    html_bytes: bytes,
    quarter_end: dt.date,
) -> Optional[Dict[str, Any]]:
    if not html_bytes:
        return None
    try:
        html = html_bytes.decode("utf-8", errors="ignore")
    except Exception:
        html = ""
    tables = read_html_tables_any(html_bytes)
    if not tables:
        return None
    scale = _detect_scale_from_text(html)

    def _norm_label(s: str) -> str:
        s = s.lower()
        s = re.sub(r"[^a-z0-9]+", " ", s)
        return re.sub(r"\s+", " ", s).strip()

    def _is_header_like(s: str) -> bool:
        if not s:
            return True
        return bool(re.search(r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec).*\d{4}", s, re.I))

    best = None
    best_score = -1
    for t in tables:
        if t is None or t.empty or t.shape[1] < 2:
            continue
        label_col = 0
        col_dates = _parse_header_dates_from_table(t)
        col_idx = None
        for idx, d in col_dates.items():
            if d == quarter_end:
                col_idx = idx
                break
        if col_idx is None:
            continue
        # If header dates occupy the label column or "$" column, shift to the next numeric column.
        if t.shape[1] > col_idx + 1:
            sample_start = 1 if t.shape[0] > 1 else 0
            sample_end = min(t.shape[0], 6) if t.shape[0] > 1 else t.shape[0]
            sample_len = max(0, sample_end - sample_start)
            numeric_here = 0
            numeric_next = 0
            for rr in range(sample_start, sample_end):
                if coerce_number(t.iat[rr, col_idx]) is not None:
                    numeric_here += 1
                if coerce_number(t.iat[rr, col_idx + 1]) is not None:
                    numeric_next += 1
            if numeric_here == 0 and numeric_next >= max(1, sample_len // 2):
                col_idx = col_idx + 1

        norm_labels: List[str] = []
        for rr in range(t.shape[0]):
            label = str(t.iat[rr, label_col]).strip()
            if not label or label.lower() == "nan":
                continue
            norm_labels.append(_norm_label(label))
        has_assets = any("total assets" in lab for lab in norm_labels)
        has_liab = any("total liabilities" in lab for lab in norm_labels) or any("total liabilities and" in lab for lab in norm_labels)
        if not (has_assets and has_liab):
            continue

        dep_cur = dep_non = None
        fin_cur = fin_non = fin_tot = None
        lease_cur = lease_non = None
        lease_fin_cur = lease_fin_non = None
        debt_cur = debt_non = debt_lt = None
        equity = None
        goodwill = None
        intang = None
        restricted_cash = None
        prepaid_other = None
        derivative_asset = None
        derivative_liab = None
        rou_asset = None
        deferred_tax = None
        st_notes_borrowings = None
        carbon_equipment_liab = None
        other_liabilities = None
        values: Dict[str, float] = {}
        labels: Dict[str, str] = {}

        for rr in range(t.shape[0]):
            label_raw = str(t.iat[rr, label_col]).strip()
            if not label_raw or label_raw.lower() == "nan":
                continue
            label_norm = _norm_label(label_raw)
            if not label_norm or _is_header_like(label_norm):
                continue
            v = coerce_number(t.iat[rr, col_idx])
            if v is None:
                continue
            v = float(v) * scale

            if "customer deposits" in label_norm:
                if "noncurrent" in label_norm or "non current" in label_norm:
                    dep_non = v
                    labels["bank_deposits_noncurrent"] = label_raw
                else:
                    dep_cur = v
                    labels["bank_deposits_current"] = label_raw
                continue
            if "finance receivables" in label_norm:
                if "short term" in label_norm or "short-term" in label_norm or "current" in label_norm:
                    fin_cur = v
                    labels["bank_finance_receivables_current"] = label_raw
                    continue
                if "long term" in label_norm or "long-term" in label_norm or "noncurrent" in label_norm or "non current" in label_norm:
                    fin_non = v
                    labels["bank_finance_receivables_noncurrent"] = label_raw
                    continue
                # total finance receivables
                if fin_cur is None and fin_non is None:
                    fin_tot = v
                    labels["bank_finance_receivables_total"] = label_raw
                    continue
            if "short term finance receivables" in label_norm or "short-term finance receivables" in label_norm:
                fin_cur = v
                labels["bank_finance_receivables_current"] = label_raw
                continue
            if "long term finance receivables" in label_norm or "long-term finance receivables" in label_norm:
                fin_non = v
                labels["bank_finance_receivables_noncurrent"] = label_raw
                continue
            if "current operating lease liabilities" in label_norm:
                lease_cur = v
                labels["lease_liabilities_current"] = label_raw
                continue
            if "noncurrent operating lease liabilities" in label_norm or "non current operating lease liabilities" in label_norm:
                lease_non = v
                labels["lease_liabilities_noncurrent"] = label_raw
                continue
            if "current finance lease liabilities" in label_norm:
                lease_fin_cur = v
                labels["lease_finance_liabilities_current"] = label_raw
                continue
            if "noncurrent finance lease liabilities" in label_norm or "non current finance lease liabilities" in label_norm:
                lease_fin_non = v
                labels["lease_finance_liabilities_noncurrent"] = label_raw
                continue
            if (
                "current portion of long-term debt" in label_norm
                or "current portion of long term debt" in label_norm
                or "current maturities of long-term debt" in label_norm
                or "current maturities of long term debt" in label_norm
            ):
                debt_cur = v
                labels["debt_core_current"] = label_raw
                continue
            if (
                "long-term debt, net of current portion" in label_norm
                or "long term debt, net of current portion" in label_norm
                or "long-term debt, less current maturities" in label_norm
                or "long term debt, less current maturities" in label_norm
                or "long-term debt noncurrent" in label_norm
                or "long term debt noncurrent" in label_norm
            ):
                debt_non = v
                labels["debt_core_noncurrent"] = label_raw
                continue
            if label_norm.startswith("long term debt") or label_norm.startswith("long-term debt"):
                debt_lt = v
                labels["debt_core_longterm"] = label_raw
                continue
            if (
                label_norm.startswith("total stockholders")
                or label_norm.startswith("total shareholders")
                or label_norm.startswith("total equity")
            ) and "liabilities and" not in label_norm:
                equity = v
                labels["total_equity"] = label_raw
                continue
            if label_norm.startswith("restricted cash"):
                restricted_cash = v
                labels["restricted_cash"] = label_raw
                continue
            if "prepaid expenses and other current assets" in label_norm or "prepaid expenses and other current" in label_norm or label_norm == "prepaid expenses and other":
                prepaid_other = v
                labels["prepaid_other_current_assets"] = label_raw
                continue
            if label_norm.startswith("other current assets") and prepaid_other is None:
                prepaid_other = v
                labels["prepaid_other_current_assets"] = label_raw
                continue
            if "derivative financial instruments" in label_norm or label_norm.startswith("derivative assets") or label_norm.startswith("derivative liabilities"):
                if "liabil" in label_norm:
                    derivative_liab = v
                    labels["derivative_liabilities"] = label_raw
                else:
                    derivative_asset = v
                    labels["derivative_assets"] = label_raw
                continue
            if "operating lease right of use assets" in label_norm or "operating lease right-of-use assets" in label_norm:
                rou_asset = v
                labels["operating_lease_rou_assets"] = label_raw
                continue
            if "deferred income taxes net" in label_norm or "deferred tax assets net" in label_norm or "deferred income tax assets net" in label_norm:
                deferred_tax = v
                labels["deferred_income_taxes_net"] = label_raw
                continue
            if label_norm.startswith("goodwill"):
                goodwill = v
                labels["goodwill"] = label_raw
                continue
            if label_norm.startswith("intangible assets"):
                intang = v
                labels["intangibles"] = label_raw
                continue
            if "short term notes payable and other borrowings" in label_norm or "short-term notes payable and other borrowings" in label_norm:
                st_notes_borrowings = v
                labels["short_term_notes_payable_and_other_borrowings"] = label_raw
                continue
            if "carbon equipment liabilities" in label_norm:
                carbon_equipment_liab = v
                labels["carbon_equipment_liabilities"] = label_raw
                continue
            if label_norm.startswith("other liabilities") or label_norm.startswith("other long term liabilities") or label_norm.startswith("other long-term liabilities"):
                if "current" not in label_norm:
                    other_liabilities = v
                    labels["other_liabilities"] = label_raw
                    continue

        if dep_cur is not None or dep_non is not None:
            if dep_cur is not None and dep_non is not None:
                values["bank_deposits"] = dep_cur + dep_non
                labels["bank_deposits"] = "customer deposits (current + noncurrent)"
            else:
                values["bank_deposits"] = dep_cur if dep_cur is not None else dep_non
                labels["bank_deposits"] = "customer deposits (partial)"
        if fin_cur is not None or fin_non is not None:
            if fin_cur is not None and fin_non is not None:
                values["bank_finance_receivables"] = fin_cur + fin_non
                labels["bank_finance_receivables"] = "finance receivables (short + long)"
            else:
                values["bank_finance_receivables"] = fin_cur if fin_cur is not None else fin_non
                labels["bank_finance_receivables"] = "finance receivables (partial)"
        elif fin_tot is not None:
            values["bank_finance_receivables"] = fin_tot
            labels["bank_finance_receivables"] = "finance receivables (total)"

        lease_parts = [x for x in [lease_cur, lease_non, lease_fin_cur, lease_fin_non] if x is not None]
        if lease_parts:
            values["lease_liabilities"] = float(sum(lease_parts))
            labels["lease_liabilities"] = "lease liabilities (operating + finance)"
        if lease_cur is not None:
            values["operating_lease_current_liabilities"] = float(lease_cur)
            labels["operating_lease_current_liabilities"] = labels.get("lease_liabilities_current", "current operating lease liabilities")
        if lease_non is not None:
            values["operating_lease_long_term_liabilities"] = float(lease_non)
            labels["operating_lease_long_term_liabilities"] = labels.get("lease_liabilities_noncurrent", "noncurrent operating lease liabilities")

        debt_val, debt_choice = choose_total_debt(
            noncurrent=float(debt_non) if debt_non is not None else None,
            current=float(debt_cur) if debt_cur is not None else None,
            longtermdebt=float(debt_lt) if debt_lt is not None else None,
        )
        if debt_cur is not None:
            values["current_maturities_of_long_term_debt"] = float(debt_cur)
            labels["current_maturities_of_long_term_debt"] = labels.get("debt_core_current", "current maturities of long-term debt")
        if debt_val is not None:
            values["debt_core"] = float(debt_val)
            if debt_choice == "noncurrent_plus_current":
                labels["debt_core"] = "noncurrent long-term debt + current portion"
            elif debt_choice == "longterm_includes_current_or_total":
                labels["debt_core"] = "long-term debt (single line)"
            elif debt_choice == "noncurrent_only":
                labels["debt_core"] = "noncurrent long-term debt"
            else:
                labels["debt_core"] = "current debt (fallback)"

        if equity is not None:
            values["total_equity"] = float(equity)
            labels["total_equity"] = labels.get("total_equity", "total equity")
        if restricted_cash is not None:
            values["restricted_cash"] = float(restricted_cash)
            labels["restricted_cash"] = labels.get("restricted_cash", "restricted cash")
        if prepaid_other is not None:
            values["prepaid_other_current_assets"] = float(prepaid_other)
            labels["prepaid_other_current_assets"] = labels.get("prepaid_other_current_assets", "prepaid expenses and other current assets")
        if derivative_asset is not None:
            values["derivative_assets"] = float(derivative_asset)
            labels["derivative_assets"] = labels.get("derivative_assets", "derivative financial instruments (asset)")
        if derivative_liab is not None:
            values["derivative_liabilities"] = float(derivative_liab)
            labels["derivative_liabilities"] = labels.get("derivative_liabilities", "derivative financial instruments (liability)")
        if rou_asset is not None:
            values["operating_lease_rou_assets"] = float(rou_asset)
            labels["operating_lease_rou_assets"] = labels.get("operating_lease_rou_assets", "operating lease right-of-use assets")
        if deferred_tax is not None:
            values["deferred_income_taxes_net"] = float(deferred_tax)
            labels["deferred_income_taxes_net"] = labels.get("deferred_income_taxes_net", "deferred income taxes, net")
        if goodwill is not None:
            values["goodwill"] = float(goodwill)
            labels["goodwill"] = labels.get("goodwill", "goodwill")
        if intang is not None:
            values["intangibles"] = float(intang)
            labels["intangibles"] = labels.get("intangibles", "intangible assets")
        if st_notes_borrowings is not None:
            values["short_term_notes_payable_and_other_borrowings"] = float(st_notes_borrowings)
            labels["short_term_notes_payable_and_other_borrowings"] = labels.get("short_term_notes_payable_and_other_borrowings", "short-term notes payable and other borrowings")
        if carbon_equipment_liab is not None:
            values["carbon_equipment_liabilities"] = float(carbon_equipment_liab)
            labels["carbon_equipment_liabilities"] = labels.get("carbon_equipment_liabilities", "carbon equipment liabilities")
        if other_liabilities is not None:
            values["other_liabilities"] = float(other_liabilities)
            labels["other_liabilities"] = labels.get("other_liabilities", "other liabilities")

        if values:
            score = len(values)
            if score > best_score:
                best_score = score
                best = {"values": values, "labels": labels, "scale": scale, "source": "tier2_table"}

    return best


def build_balance_sheet_fallback_table(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_quarters: int,
    *,
    target_quarters: Optional[set[dt.date]] = None,
    filing_inventory: Optional[Dict[str, Any]] = None,
    filing_runtime_cache: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[dt.date, Dict[str, Any]], List[Dict[str, Any]]]:
    out: Dict[dt.date, Dict[str, Any]] = {}
    audit_rows: List[Dict[str, Any]] = []
    seen_accn: set[str] = set()
    target_years = _build_target_years_from_quarters(target_quarters)
    filing_inventory, filing_runtime_cache = _ensure_primary_filing_inventory(
        sec,
        submissions,
        filing_inventory=filing_inventory,
        filing_runtime_cache=filing_runtime_cache,
        target_years=target_years,
    )
    for row in filing_inventory.get("rows", []) or []:
        form = row.get("form")
        if form not in ("10-Q", "10-Q/A", "10-K", "10-K/A"):
            continue
        accn = row.get("accn")
        if accn in seen_accn:
            continue
        doc = row.get("primary_doc")
        if not doc:
            continue
        q_end = row.get("report_date") or row.get("filing_date")
        if q_end is None:
            continue
        if target_quarters is not None and q_end not in target_quarters:
            continue
        accn_nd = str(row.get("accn_nd") or "")
        html_bytes = _load_primary_filing_document_bytes(sec, cik_int, accn_nd, str(doc), filing_runtime_cache)
        if html_bytes is None:
            continue
        seen_accn.add(str(accn))
        result = _extract_balance_sheet_from_html(html_bytes, q_end)
        if not result:
            continue
        payload = dict(result)
        payload.update({"accn": accn, "doc": doc, "form": form, "filed": row.get("filing_date_raw")})
        if q_end in out and len(out[q_end].get("values", {})) >= len(result.get("values", {})):
            continue
        out[q_end] = payload
        audit_rows.append({
            "quarter": q_end,
            "source": "tier2_table",
            "accn": accn,
            "doc": doc,
            "note": "balance sheet fallback from 10-Q/10-K table",
        })

    qs = sorted(out.keys())[-max_quarters:]
    out = {q: out[q] for q in qs}
    return out, audit_rows


def build_income_statement_fallback_ex99_ocr(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_quarters: int,
    *,
    ticker: Optional[str] = None,
    target_quarters: Optional[set[dt.date]] = None,
    ex99_inventory: Optional[Dict[str, Any]] = None,
    ex99_runtime_cache: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[dt.date, Dict[str, Any]], List[Dict[str, Any]]]:
    out: Dict[dt.date, Dict[str, Any]] = {}
    audit_rows: List[Dict[str, Any]] = []
    rules = get_income_statement_rules(ticker)
    target_quarters = set(target_quarters) if target_quarters else None
    target_years = _build_target_years_from_quarters(target_quarters)
    ex99_inventory, ex99_runtime_cache = _ensure_ex99_inventory(
        sec,
        cik_int,
        submissions,
        ex99_inventory=ex99_inventory,
        ex99_runtime_cache=ex99_runtime_cache,
        target_years=target_years,
    )

    def _accept_q(q_end: Optional[dt.date]) -> bool:
        if q_end is None:
            return False
        if not _is_quarter_end(q_end):
            return False
        if target_quarters is None:
            return True
        return q_end in target_quarters

    for row in ex99_inventory.get("eight_k_rows", []) or []:
        if target_quarters is not None and not target_quarters:
            break
        accn = row.get("accn")
        accn_nd = str(row.get("accn_nd") or "")
        if not accn_nd:
            continue
        rdate = row.get("report_date_raw")
        fdate = row.get("filing_date_raw")
        if not _dates_match_target_years(row.get("report_date"), row.get("filing_date"), target_years):
            continue
        exdocs = list(row.get("exdocs") or [])
        idx = _load_ex99_accession_index(sec, cik_int, accn_nd, ex99_runtime_cache)
        if idx is None:
            continue
        if not exdocs:
            imgs = _load_ex99_index_images(sec, cik_int, accn_nd, idx, ex99_runtime_cache)
            if imgs:
                try:
                    ocr_txt = sec.ocr_html_assets(
                        accn_nd,
                        None,
                        context={
                            "doc": "index_images",
                            "quarter": None,
                            "purpose": "is_ocr",
                            "report_date": rdate,
                            "filing_date": fdate,
                            "save_text": True,
                        },
                    )
                except Exception:
                    ocr_txt = ""
                if ocr_txt:
                    q_end = infer_quarter_end_from_text(ocr_txt) or parse_date(rdate) or parse_date(fdate)
                    if not _is_quarter_end(q_end):
                        q_end = None
                    if _accept_q(q_end):
                        result = _extract_income_statement_from_text(ocr_txt, q_end, rules=rules)
                        if result:
                            result["accn"] = accn
                            result["doc"] = "index_images"
                            if result.get("tokens"):
                                sec.ocr_log_rows.append({
                                    "accn": accn,
                                    "doc": "index_images",
                                    "quarter": q_end,
                                    "purpose": "is_ocr_match",
                                    "status": "ok",
                                    "ocr_tokens": "; ".join([f"{k}: {v}" for k, v in result["tokens"].items() if v]),
                                })
                            out[q_end] = result
                            if target_quarters is not None:
                                target_quarters.discard(q_end)
                            audit_rows.append({
                                "quarter": q_end,
                                "source": "tier3_ex99_ocr",
                                "accn": accn,
                                "doc": "index_images",
                                "note": "income statement OCR from image-only 8-K",
                            })
            continue
        for fn in exdocs[:6]:
            b = _load_ex99_document_bytes(sec, cik_int, accn_nd, fn, ex99_runtime_cache)
            if b is None:
                continue
            try:
                sec.download_html_assets(cik_int, accn_nd, b)
            except Exception:
                pass
            try:
                txt = sec.ocr_html_assets(
                    accn_nd,
                    b,
                    context={"doc": fn, "purpose": "is_ocr", "report_date": rdate, "filing_date": fdate, "save_text": True},
                )
            except Exception:
                txt = ""
            q_end = infer_quarter_end_from_text(txt) or parse_date(rdate) or parse_date(fdate)
            if q_end is None or not _accept_q(q_end):
                continue
            result = _extract_income_statement_from_text(txt, q_end, rules=rules)
            if not result:
                continue
            result["accn"] = accn
            result["doc"] = fn
            if result.get("tokens"):
                sec.ocr_log_rows.append({
                    "accn": accn,
                    "doc": fn,
                    "quarter": q_end,
                    "purpose": "is_ocr_match",
                    "status": "ok",
                    "ocr_tokens": "; ".join([f"{k}: {v}" for k, v in result["tokens"].items() if v]),
                })
            out[q_end] = result
            if target_quarters is not None:
                target_quarters.discard(q_end)
            audit_rows.append({
                "quarter": q_end,
                "source": "tier3_ex99_ocr",
                "accn": accn,
                "doc": fn,
                "note": "income statement OCR from EX-99",
            })
            break

    # Legacy cache scan for older EX-99 docs (OCR)
    scanned = 0
    for entry in ex99_inventory.get("legacy_docs", []) or []:
        if target_quarters is not None and not target_quarters:
            break
        if scanned >= 120:
            break
        if not entry.get("is_html_like"):
            continue
        accn_nd = entry.get("accn_nd")
        fn = entry.get("doc") or entry.get("name")
        path_in = entry.get("path")
        if not accn_nd or not isinstance(path_in, Path):
            continue
        b = _load_legacy_ex99_document_bytes(path_in, ex99_runtime_cache)
        if b is None:
            continue
        try:
            sec.download_html_assets(cik_int, accn_nd, b)
        except Exception:
            pass
        try:
            txt = sec.ocr_html_assets(
                accn_nd,
                b,
                context={"doc": fn, "purpose": "is_ocr_legacy", "save_text": False},
            )
        except Exception:
            txt = ""
        scanned += 1
        if not txt:
            continue
        q_end = infer_quarter_end_from_text(txt)
        if not _accept_q(q_end):
            continue
        result = _extract_income_statement_from_text(txt, q_end, rules=rules)
        if not result:
            continue
        result["accn"] = accn_nd
        result["doc"] = fn
        out[q_end] = result
        if target_quarters is not None:
            target_quarters.discard(q_end)
        audit_rows.append({
            "quarter": q_end,
            "source": "tier3_ex99_ocr",
            "accn": accn_nd,
            "doc": fn,
            "note": "income statement OCR from cached EX-99",
        })

    if out:
        qs = sorted(out.keys())[-max_quarters:]
        out = {q: out[q] for q in qs}
    return out, audit_rows


def build_balance_sheet_fallback_ex99_ocr(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_quarters: int,
    *,
    target_quarters: Optional[set[dt.date]] = None,
    ex99_inventory: Optional[Dict[str, Any]] = None,
    ex99_runtime_cache: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[dt.date, Dict[str, Any]], List[Dict[str, Any]]]:
    out: Dict[dt.date, Dict[str, Any]] = {}
    audit_rows: List[Dict[str, Any]] = []
    target_quarters = set(target_quarters) if target_quarters else None
    target_years = _build_target_years_from_quarters(target_quarters)
    ex99_inventory, ex99_runtime_cache = _ensure_ex99_inventory(
        sec,
        cik_int,
        submissions,
        ex99_inventory=ex99_inventory,
        ex99_runtime_cache=ex99_runtime_cache,
        target_years=target_years,
    )

    def _accept_q(q_end: Optional[dt.date]) -> bool:
        if q_end is None:
            return False
        if not _is_quarter_end(q_end):
            return False
        if target_quarters is None:
            return True
        return q_end in target_quarters

    for row in ex99_inventory.get("eight_k_rows", []) or []:
        if target_quarters is not None and not target_quarters:
            break
        accn = row.get("accn")
        accn_nd = str(row.get("accn_nd") or "")
        if not accn_nd:
            continue
        rdate = row.get("report_date_raw")
        fdate = row.get("filing_date_raw")
        if not _dates_match_target_years(row.get("report_date"), row.get("filing_date"), target_years):
            continue
        exdocs = list(row.get("exdocs") or [])
        idx = _load_ex99_accession_index(sec, cik_int, accn_nd, ex99_runtime_cache)
        if idx is None:
            continue
        if not exdocs:
            imgs = _load_ex99_index_images(sec, cik_int, accn_nd, idx, ex99_runtime_cache)
            if imgs:
                try:
                    txt = sec.ocr_html_assets(
                        accn_nd,
                        None,
                        context={"doc": "index_images", "purpose": "bs_ocr", "report_date": rdate, "filing_date": fdate, "save_text": True},
                    )
                except Exception:
                    txt = ""
                q_end = infer_quarter_end_from_text(txt) or parse_date(rdate) or parse_date(fdate)
                if _accept_q(q_end):
                    result = _extract_balance_sheet_from_text(txt, q_end)
                    if result:
                        result["accn"] = accn
                        result["doc"] = "index_images"
                        if result.get("tokens"):
                            sec.ocr_log_rows.append({
                                "accn": accn,
                                "doc": "index_images",
                                "quarter": q_end,
                                "purpose": "bs_ocr_match",
                                "status": "ok",
                                "ocr_tokens": "; ".join([f"{k}: {v}" for k, v in result["tokens"].items() if v]),
                            })
                        out[q_end] = result
                        if target_quarters is not None:
                            target_quarters.discard(q_end)
                        audit_rows.append({
                            "quarter": q_end,
                            "source": "tier3_ex99_ocr",
                            "accn": accn,
                            "doc": "index_images",
                            "note": "balance sheet OCR from image-only 8-K",
                        })
            continue
        for fn in exdocs[:6]:
            b = _load_ex99_document_bytes(sec, cik_int, accn_nd, fn, ex99_runtime_cache)
            if b is None:
                continue
            try:
                sec.download_html_assets(cik_int, accn_nd, b)
            except Exception:
                pass
            try:
                txt = sec.ocr_html_assets(
                    accn_nd,
                    b,
                    context={"doc": fn, "purpose": "bs_ocr", "report_date": rdate, "filing_date": fdate, "save_text": True},
                )
            except Exception:
                txt = ""
            q_end = infer_quarter_end_from_text(txt) or parse_date(rdate) or parse_date(fdate)
            if q_end is None or not _accept_q(q_end):
                continue
            result = _extract_balance_sheet_from_text(txt, q_end)
            if not result:
                continue
            if result.get("tokens"):
                sec.ocr_log_rows.append({
                    "accn": accn,
                    "doc": fn,
                    "quarter": q_end,
                    "purpose": "bs_ocr_match",
                    "status": "ok",
                    "ocr_tokens": "; ".join([f"{k}: {v}" for k, v in result["tokens"].items() if v]),
                })
            out[q_end] = result
            if target_quarters is not None:
                target_quarters.discard(q_end)
            audit_rows.append({
                "quarter": q_end,
                "source": "tier3_ex99_ocr",
                "accn": accn,
                "doc": fn,
                "note": "balance sheet OCR from EX-99",
            })
            break

    # Legacy cache scan for older EX-99 docs (OCR)
    scanned = 0
    for entry in ex99_inventory.get("legacy_docs", []) or []:
        if target_quarters is not None and not target_quarters:
            break
        if scanned >= 120:
            break
        if not entry.get("is_html_like"):
            continue
        accn_nd = entry.get("accn_nd")
        fn = entry.get("doc") or entry.get("name")
        path_in = entry.get("path")
        if not accn_nd or not isinstance(path_in, Path):
            continue
        b = _load_legacy_ex99_document_bytes(path_in, ex99_runtime_cache)
        if b is None:
            continue
        try:
            sec.download_html_assets(cik_int, accn_nd, b)
        except Exception:
            pass
        try:
            txt = sec.ocr_html_assets(
                accn_nd,
                b,
                context={"doc": fn, "purpose": "bs_ocr_legacy", "save_text": False},
            )
        except Exception:
            txt = ""
        scanned += 1
        if not txt:
            continue
        q_end = infer_quarter_end_from_text(txt)
        if not _accept_q(q_end):
            continue
        result = _extract_balance_sheet_from_text(txt, q_end)
        if not result:
            continue
        result["accn"] = accn_nd
        result["doc"] = fn
        out[q_end] = result
        if target_quarters is not None:
            target_quarters.discard(q_end)
        audit_rows.append({
            "quarter": q_end,
            "source": "tier3_ex99_ocr",
            "accn": accn_nd,
            "doc": fn,
            "note": "balance sheet OCR from cached EX-99",
        })

    if out:
        qs = sorted(out.keys())[-max_quarters:]
        out = {q: out[q] for q in qs}
    return out, audit_rows


def _extract_eps_shares_from_html(
    html_bytes: bytes,
    quarter_end: dt.date,
) -> Optional[Dict[str, Any]]:
    html = html_bytes.decode("utf-8", errors="ignore")
    eps_diluted_re = re.compile(r"weighted[- ]?average\s+shares.*dilut", re.I)
    eps_diluted_tokens = ["weighted", "average", "share", "dilut"]
    scale = 1000.0 if re.search(r"in\s+thousands", html, re.I) else 1.0
    tables = read_html_tables_any(html_bytes)
    for t in tables:
        if t is None or t.empty:
            continue
        t2 = t.copy()
        header_text = " ".join([str(c) for c in t2.columns]).lower()
        body_text = _table_head_text(t2, 20)
        table_text = header_text + " " + body_text
        if "earnings per share" not in table_text and "eps" not in table_text:
            continue
        if "three months ended" not in table_text and "quarter" not in table_text:
            continue
        # pick column that matches quarter_end
        col_dates = _parse_header_dates_from_table(t2)
        col_idx_pos = None
        if col_dates:
            for i, d in col_dates.items():
                if d == quarter_end:
                    col_idx_pos = i
                    break
        if col_idx_pos is None and len(t2) >= 3:
            # Fallback: detect "Three Months Ended" header with year row beneath
            row1 = [str(x) for x in t2.iloc[1].tolist()]
            row2 = [str(x) for x in t2.iloc[2].tolist()]
            for i in range(min(len(row1), len(t2.columns))):
                if "three months ended" in row1[i].lower() and str(quarter_end.year) in row2[i]:
                    col_idx_pos = i
                    break
        if col_idx_pos is None:
            continue
        # label col by alpha density
        alpha_scores = {}
        for i, c in enumerate(t2.columns):
            vals = _stringify_table_cells(t2[c].head(25).tolist())
            alpha_scores[i] = sum(1 for v in vals if re.search(r"[A-Za-z]", v))
        label_col_idx = max(alpha_scores, key=alpha_scores.get)
        rows = t2.iloc[:, [label_col_idx, col_idx_pos]].copy()
        rows.iloc[:, 0] = rows.iloc[:, 0].astype(str)

        diluted = None
        basic = None
        dil_label = ""
        bas_label = ""

        for _, r in rows.iterrows():
            label = str(r.iloc[0]).strip().lower()
            if not label or label == "nan":
                continue
            v = coerce_number(r.iloc[1])
            if v is None:
                continue
            if eps_diluted_re.search(label) or all(tok in label for tok in eps_diluted_tokens):
                diluted = float(v)
                dil_label = label
            if "weighted-average shares used in basic" in label:
                basic = float(v)
                bas_label = label

        if diluted is not None:
            diluted = float(diluted) * scale
            if diluted < 5_000_000:
                diluted = diluted * 1000.0
            if scale == 1000.0 and diluted is not None:
                diluted = round(diluted / 1000.0) * 1000.0
            if diluted is not None:
                diluted = round(diluted, 3)
            if not (0 < diluted <= 5_000_000_000):
                diluted = None
        if basic is not None:
            basic = float(basic) * scale
            if basic < 5_000_000:
                basic = basic * 1000.0
            if scale == 1000.0 and basic is not None:
                basic = round(basic / 1000.0) * 1000.0
            if basic is not None:
                basic = round(basic, 3)
            if not (0 < basic <= 5_000_000_000):
                basic = None

        if diluted is None and basic is None:
            continue

        return {
            "shares_diluted": diluted,
            "shares_basic": basic,
            "label_diluted": dil_label,
            "label_basic": bas_label,
        }

    # Fallback: parse raw HTML row if tables are not parsed well (e.g., EX-99.1)
    m = eps_diluted_re.search(html)
    if not m:
        for ln in html.splitlines():
            l = ln.strip().lower()
            if not l:
                continue
            if all(tok in l for tok in eps_diluted_tokens):
                nums = re.findall(r"(-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d{4,}(?:\.\d+)?)", l)
                if nums:
                    val = float(nums[0].replace(",", ""))
                    if val < 5_000_000:
                        val *= 1000.0
                    if 0 < val <= 5_000_000_000:
                        return {
                            "shares_diluted": val,
                            "shares_basic": None,
                            "label_diluted": f"{l} (line)",
                            "label_basic": "",
                        }
        return None
    # look back for "three months ended" header
    back = html[max(0, m.start() - 2000):m.start()]
    years = re.findall(r"(20\\d{2})", back)
    years = [int(y) for y in years][-4:]  # keep recent
    # extract numbers after the label
    after = html[m.start():m.start() + 2000]
    nums = re.findall(r"(-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d{4,}(?:\.\d+)?)", after)
    if not nums:
        return None
    vals = [float(n.replace(",", "")) for n in nums]
    # if "in thousands" appears anywhere (header may be far from row), scale
    if re.search(r"in\s+thousands", back + after, re.I) or re.search(r"in\s+thousands", html, re.I):
        scale = 1000.0
    else:
        scale = 1.0
    vals = [v * scale for v in vals]
    # heuristic: first two numbers correspond to "three months ended" (current, prior)
    diluted = None
    if len(vals) >= 2:
        if years:
            # assume first year in years list is current year
            cur_year = quarter_end.year
            if cur_year in years:
                # pick first number as current year
                diluted = vals[0]
            else:
                diluted = vals[0]
        else:
            diluted = vals[0]
    elif vals:
        diluted = vals[0]
    if diluted is None or not (0 < diluted <= 5_000_000_000):
        return None
    if scale == 1000.0:
        diluted = round(diluted / 1000.0) * 1000.0
    diluted = round(diluted, 3)
    return {
        "shares_diluted": diluted,
        "shares_basic": None,
        "label_diluted": "weighted-average shares used in diluted earnings per share (html)",
        "label_basic": "",
    }


def _extract_eps_shares_from_text(
    text: str,
    quarter_end: Optional[dt.date],
) -> Optional[Dict[str, Any]]:
    if not text:
        return None
    low = text.lower()
    if "weighted-average" not in low and "weighted average" not in low:
        return None
    in_thousands = bool(re.search(r"in\s+thousands", low))
    lines = [re.sub(r"\s+", " ", ln).strip() for ln in text.splitlines() if ln.strip()]
    target_idx = None
    for i, ln in enumerate(lines):
        if re.search(r"weighted[- ]?average\s+shares.*dilut", ln, re.I):
            target_idx = i
            break
    if target_idx is None:
        return None
    cand = [lines[target_idx]]
    if target_idx + 1 < len(lines):
        cand.append(lines[target_idx + 1])
    blob = " ".join(cand)
    nums = re.findall(r"(-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d{4,}(?:\.\d+)?)", blob)
    if not nums:
        return None
    vals = [float(n.replace(",", "")) for n in nums]
    # Detect scale: if values look like "in thousands", apply scale only when values are small
    if in_thousands and all(v < 5_000_000 for v in vals):
        vals = [v * 1000.0 for v in vals]
    # Drop extreme outlier if present (OCR often merges columns)
    if len(vals) >= 3:
        med = sorted(vals)[len(vals) // 2]
        if med > 0:
            keep = [v for v in vals if v <= med * 2.5]
            if keep:
                vals = keep
    # Heuristic: first remaining number is current quarter
    val = vals[0]
    if val < 5_000_000:
        val *= 1000.0
    if not (0 < val <= 5_000_000_000):
        return None
    if in_thousands:
        val = round(val / 1000.0) * 1000.0
    val = round(val, 3)
    return {
        "shares_diluted": float(val),
        "shares_basic": None,
        "label_diluted": blob[:240],
        "label_basic": "",
    }


def _extract_cash_taxes_from_html(
    html_bytes: bytes,
    quarter_end: dt.date,
    *,
    period_hint: str = "3M",
) -> Optional[Dict[str, Any]]:
    html = html_bytes.decode("utf-8", errors="ignore")
    scale = _detect_scale_from_text(html)
    tables = read_html_tables_any(html_bytes)
    label_re = re.compile(
        r"cash\s+income\s+tax(?:es)?\s*(?:\(refunds\)\s*)?payments?(?:\s*\(refunds\))?,?\s*net",
        re.IGNORECASE,
    )
    if period_hint == "3M":
        period_tokens = ["three months", "three-months"]
    elif period_hint == "6M":
        period_tokens = ["six months", "six-months"]
    elif period_hint == "9M":
        period_tokens = ["nine months", "nine-months"]
    else:
        period_tokens = []
    def _year_positions(t2: pd.DataFrame) -> List[Tuple[int, int]]:
        best: List[Tuple[int, int]] = []
        for ridx in range(min(3, len(t2))):
            row = [str(x) for x in t2.iloc[ridx].tolist()]
            yrs: List[Tuple[int, int]] = []
            for ci, cell in enumerate(row):
                m = re.search(r"(19|20)\d{2}", cell)
                if m:
                    yrs.append((ci, int(m.group(0))))
            if len(yrs) > len(best):
                best = yrs
        return best
    for t in tables:
        if t is None or t.empty:
            continue
        t2 = t.copy()
        table_text = _table_head_text(t2, 25)
        table_text = table_text.replace("\xa0", " ")
        if "tax" not in table_text:
            continue

        # pick column that matches quarter_end and period_hint
        col_dates = _parse_header_dates_from_table(t2)
        col_idx = None
        col_texts = []
        if not t2.empty:
            row0 = [str(x) for x in t2.columns]
            row1 = [str(x) for x in t2.iloc[0].tolist()] if len(t2) > 0 else [""] * len(row0)
            row2 = [str(x) for x in t2.iloc[1].tolist()] if len(t2) > 1 else [""] * len(row0)
            for i in range(len(row0)):
                col_texts.append(" ".join([row0[i], row1[i], row2[i]]).lower())
        if col_dates:
            candidates = [i for i, d in col_dates.items() if d == quarter_end]
            if candidates and period_tokens:
                prefer = []
                for i in candidates:
                    if i < len(col_texts) and any(tok in col_texts[i] for tok in period_tokens):
                        prefer.append(i)
                if prefer:
                    col_idx = t2.columns[prefer[0]]
            if col_idx is None and candidates:
                col_idx = t2.columns[candidates[0]]
        if col_idx is None and len(t2) >= 3:
            row1 = [str(x) for x in t2.iloc[1].tolist()]
            row2 = [str(x) for x in t2.iloc[2].tolist()]
            for i in range(min(len(row1), len(t2.columns))):
                r1 = row1[i].lower()
                if period_hint == "3M" and "three months ended" in r1 and str(quarter_end.year) in row2[i]:
                    col_idx = t2.columns[i]
                    break
                if period_hint == "6M" and "six months ended" in r1 and str(quarter_end.year) in row2[i]:
                    col_idx = t2.columns[i]
                    break
                if period_hint == "9M" and ("nine months ended" in r1 or "nine-months ended" in r1) and str(quarter_end.year) in row2[i]:
                    col_idx = t2.columns[i]
                    break
        top_text = _table_head_text(t2, 3)
        top_text = re.sub(r"\s+", " ", top_text.replace("\xa0", " ")).strip()

        year_pos = _year_positions(t2)
        year_order = [y for _, y in sorted(year_pos, key=lambda x: x[0])]

        for _, r in t2.iterrows():
            row_vals = [str(x) for x in r.tolist()]
            row_text = " ".join(row_vals).lower()
            if not label_re.search(row_text):
                continue
            label = ""
            best_alpha = -1
            for vtxt in row_vals:
                score = sum(1 for ch in vtxt if ch.isalpha())
                if score > best_alpha:
                    best_alpha = score
                    label = vtxt.strip().lower()
            if not label:
                label = row_text[:120]
            v = None
            if col_idx is not None:
                v = coerce_number(r[col_idx])
            if v is None:
                # Fallback: map year labels to numeric values by order
                nums: List[float] = []
                for c in t2.columns:
                    n = coerce_number(r[c])
                    if n is not None:
                        nums.append(float(n))
                if year_order and nums:
                    if len(nums) > len(year_order):
                        if period_hint in ("6M", "9M"):
                            nums = nums[-len(year_order):]
                        else:
                            nums = nums[:len(year_order)]
                if year_order and nums and quarter_end.year in year_order:
                    yi = year_order.index(quarter_end.year)
                    if yi < len(nums):
                        v = nums[yi]
            if v is None:
                continue
            val = float(v) * scale
            return {"value": val, "label": label}

    return None


def _cash_tax_label_present(html_bytes: bytes) -> bool:
    text = html_bytes.decode("utf-8", errors="ignore").lower()
    if "cash" not in text or "tax" not in text:
        return False
    return bool(re.search(r"cash\s+income\s+tax", text) and re.search(r"payments?", text))


def _extract_text_from_pdf_bytes(data: bytes, quiet_pdf_warnings: bool = True) -> str:
    try:
        import pdfplumber  # type: ignore
    except Exception:
        return ""
    try:
        with silence_pdfminer_warnings(enabled=quiet_pdf_warnings):
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                texts = [page.extract_text() or "" for page in pdf.pages]
        return "\n".join([t for t in texts if t])
    except Exception:
        return ""


def build_shares_outstanding_fallback(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_quarters: int,
    *,
    filing_inventory: Optional[Dict[str, Any]] = None,
    filing_runtime_cache: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[dt.date, Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Extract period-end common shares outstanding from cover page text.
    Pattern example:
    "As of October 20, 2025, 160,918,164 shares of common stock ... were outstanding."
    """
    results: Dict[dt.date, Dict[str, Any]] = {}
    audit: List[Dict[str, Any]] = []
    scanned = 0
    filing_inventory, filing_runtime_cache = _ensure_primary_filing_inventory(
        sec,
        submissions,
        filing_inventory=filing_inventory,
        filing_runtime_cache=filing_runtime_cache,
    )
    for row in filing_inventory.get("rows", []) or []:
        form = row.get("form")
        if form not in ("10-Q", "10-Q/A", "10-K", "10-K/A"):
            continue
        accn = row.get("accn")
        doc = row.get("primary_doc")
        if not doc:
            continue
        q_end = row.get("report_date") or row.get("filing_date")
        if not _is_quarter_end(q_end):
            q_end = _coerce_prev_quarter_end(q_end)
        if q_end is None:
            continue
        accn_nd = str(row.get("accn_nd") or "")
        html_bytes = _load_primary_filing_document_bytes(sec, cik_int, accn_nd, str(doc), filing_runtime_cache)
        if html_bytes is None:
            continue
        text = strip_html(html_bytes.decode("utf-8", errors="ignore"))
        text = html.unescape(text).replace("\xa0", " ")
        if not text:
            continue
        snippet = "\n".join(text.splitlines()[:120])
        patterns = [
            r"(?:as of|at)\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})\s*,?\s*([0-9][0-9,]+)\s+shares\s+of\s+common\s+stock.*?(?:were\s+outstanding|outstanding)",
            r"(?:as of|at)\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})[^.\n]{0,160}there\s+were\s+([0-9][0-9,]+)\s+outstanding\s+shares",
            r"number of shares of common stock.*?outstanding as of (?:close of business on )?([A-Za-z]+\s+\d{1,2},\s+\d{4})\s*:?\s*([0-9][0-9,]+)\s+shares",
            r"at\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})\s*,?\s*there were\s+([0-9][0-9,]+)\s+outstanding\s+shares",
        ]
        m = None
        for pat in patterns:
            m = re.search(pat, snippet, re.I)
            if m:
                break
        if not m:
            for pat in patterns:
                m = re.search(pat, text, re.I)
                if m:
                    break
        if not m:
            continue
        asof = parse_date(m.group(1))
        val = coerce_number(m.group(2))
        if val is None:
            continue
        results[q_end] = {
            "shares_outstanding": float(val),
            "as_of": asof,
            "accn": accn,
            "form": form,
            "filed": row.get("filing_date"),
            "doc": doc,
            "snippet": m.group(0)[:180],
        }
        audit.append({
            "metric": "shares_outstanding",
            "quarter": q_end,
            "source": "cover_page",
            "tag": "Cover page",
            "accn": accn,
            "form": form,
            "filed": row.get("filing_date"),
            "start": None,
            "end": q_end,
            "unit": "shares",
            "duration_days": None,
            "value": float(val),
            "note": f"Cover page shares outstanding (as of {asof})",
        })
        scanned += 1
        if scanned >= max_quarters:
            break
    return results, audit


def build_eps_shares_fallback(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_quarters: int,
    *,
    filing_inventory: Optional[Dict[str, Any]] = None,
    filing_runtime_cache: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[dt.date, Dict[str, Any]], List[Dict[str, Any]]]:
    out: Dict[dt.date, Dict[str, Any]] = {}
    audit_rows: List[Dict[str, Any]] = []
    seen_accn: set[str] = set()
    filing_inventory, filing_runtime_cache = _ensure_primary_filing_inventory(
        sec,
        submissions,
        filing_inventory=filing_inventory,
        filing_runtime_cache=filing_runtime_cache,
    )
    for row in filing_inventory.get("rows", []) or []:
        form = row.get("form")
        if form not in ("10-Q", "10-Q/A"):
            continue
        accn = row.get("accn")
        if accn in seen_accn:
            continue
        doc = row.get("primary_doc")
        if not doc:
            continue
        q_end = row.get("report_date") or row.get("filing_date")
        accn_nd = str(row.get("accn_nd") or "")
        html_bytes = _load_primary_filing_document_bytes(sec, cik_int, accn_nd, str(doc), filing_runtime_cache)
        if html_bytes is None:
            continue
        raw = html_bytes.decode("utf-8", errors="ignore")
        if q_end is None:
            q_end = infer_quarter_end_from_text(raw)
        if q_end is None or not _is_quarter_end(q_end):
            continue
        seen_accn.add(str(accn))
        result = _extract_eps_shares_from_html(html_bytes, q_end)
        if not result and "earnings per share" in raw.lower():
            result = _extract_eps_shares_from_text(raw, q_end)
        if not result:
            continue
        out[q_end] = result
        audit_rows.append({
            "quarter": q_end,
            "source": "tier2_eps_note",
            "accn": accn,
            "doc": doc,
            "note": "EPS note fallback from 10-Q table",
        })

    # trim to recent quarters
    if out:
        qs = sorted(out.keys())[-max_quarters:]
        out = {q: out[q] for q in qs}
    else:
        # Fallback to cached EX-99 docs if index lookups miss
        cache_dir = getattr(sec, "cache_dir", None)
        if cache_dir is not None:
            for p in Path(cache_dir).glob("doc_*ex99*.htm*"):
                try:
                    raw = p.read_text(encoding="utf-8", errors="ignore")
                except Exception:
                    continue
                txt = strip_html(raw)
                q_end = infer_quarter_end_from_text(txt)
                if q_end is None:
                    continue
                m = re.search(r"weighted[- ]?average shares.*diluted", raw, re.I)
                if not m:
                    continue
                row = raw[m.start():m.start() + 2000]
                nums = re.findall(r"(-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d{4,}(?:\.\d+)?)", row)
                if not nums:
                    continue
                val = float(nums[0].replace(",", ""))
                if re.search(r"in\s+thousands", raw, re.I):
                    val *= 1000.0
                if not (0 < val <= 5_000_000_000):
                    continue
                out[q_end] = {
                    "shares_diluted": val,
                    "shares_basic": None,
                    "label_diluted": "weighted-average shares used in diluted earnings per share (ex99 cache)",
                    "label_basic": "",
                }
            if out:
                qs = sorted(out.keys())[-max_quarters:]
                out = {q: out[q] for q in qs}
    return out, audit_rows


def build_cash_taxes_fallback_10q(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_quarters: int,
    *,
    period_hint: str = "3M",
    target_quarters: Optional[set[dt.date]] = None,
    filing_inventory: Optional[Dict[str, Any]] = None,
    filing_runtime_cache: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[dt.date, Dict[str, Any]], List[Dict[str, Any]]]:
    out: Dict[dt.date, Dict[str, Any]] = {}
    audit_rows: List[Dict[str, Any]] = []
    seen_accn: set[str] = set()
    target_quarters = set(target_quarters) if target_quarters else None
    filing_inventory, filing_runtime_cache = _ensure_primary_filing_inventory(
        sec,
        submissions,
        filing_inventory=filing_inventory,
        filing_runtime_cache=filing_runtime_cache,
        target_years=_build_target_years_from_quarters(target_quarters),
    )

    for row in filing_inventory.get("rows", []) or []:
        form = row.get("form")
        if form not in ("10-Q", "10-Q/A"):
            continue
        accn = row.get("accn")
        if accn in seen_accn:
            continue
        q_end = row.get("report_date") or row.get("filing_date")
        if not _is_quarter_end(q_end):
            q_end = _coerce_prev_quarter_end(q_end)
        if q_end is None:
            continue
        if target_quarters is not None and q_end not in target_quarters:
            continue
        doc = row.get("primary_doc")
        if not doc:
            continue
        accn_nd = str(row.get("accn_nd") or "")
        html_bytes = _load_primary_filing_document_bytes(sec, cik_int, accn_nd, str(doc), filing_runtime_cache)
        if html_bytes is None:
            continue
        seen_accn.add(str(accn))
        res = _extract_cash_taxes_from_html(html_bytes, q_end, period_hint=period_hint)
        if not res:
            if _cash_tax_label_present(html_bytes):
                audit_rows.append({
                    "quarter": q_end,
                    "source": "cash_taxes_missing",
                    "accn": accn,
                    "doc": doc,
                    "note": f"Cash tax row present but not parsed ({period_hint})",
                })
            continue
        out[q_end] = {
            "value": float(res.get("value")),
            "label": res.get("label") or "",
            "accn": accn,
            "doc": doc,
            "form": form,
            "report_date": row.get("report_date_raw"),
            "filing_date": row.get("filing_date_raw"),
        }
        audit_rows.append({
            "quarter": q_end,
            "source": "tier3_cash_taxes",
            "accn": accn,
            "doc": doc,
            "note": f"Cash income tax payments (supplemental, {period_hint})",
        })
        if target_quarters is not None:
            target_quarters.discard(q_end)
        if target_quarters is not None and not target_quarters:
            break

    if out:
        qs = sorted(out.keys())[-max_quarters:]
        out = {q: out[q] for q in qs}
    return out, audit_rows


def build_eps_shares_fallback_ex99(
    sec: SecClient,
    cik_int: int,
    submissions: Dict[str, Any],
    max_quarters: int,
    *,
    target_quarters: Optional[set[dt.date]] = None,
    quiet_pdf_warnings: bool = True,
    ex99_inventory: Optional[Dict[str, Any]] = None,
    ex99_runtime_cache: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[dt.date, Dict[str, Any]], List[Dict[str, Any]]]:
    out: Dict[dt.date, Dict[str, Any]] = {}
    audit_rows: List[Dict[str, Any]] = []
    seen_accn: set[str] = set()
    target_quarters = set(target_quarters) if target_quarters else None
    target_years = _build_target_years_from_quarters(target_quarters)
    ex99_inventory, ex99_runtime_cache = _ensure_ex99_inventory(
        sec,
        cik_int,
        submissions,
        ex99_inventory=ex99_inventory,
        ex99_runtime_cache=ex99_runtime_cache,
        target_years=target_years,
    )
    accn_dates = dict(ex99_inventory.get("accn_dates", {}) or {})

    def _accept_q(q_end: Optional[dt.date]) -> bool:
        if q_end is None:
            return False
        if not _is_quarter_end(q_end):
            return False
        if target_quarters is None:
            return True
        return q_end in target_quarters

    for row in ex99_inventory.get("eight_k_rows", []) or []:
        if target_quarters is not None and not target_quarters:
            break
        accn = row.get("accn")
        if accn in seen_accn:
            continue
        seen_accn.add(accn)
        accn_nd = str(row.get("accn_nd") or "")
        if not accn_nd:
            continue
        rdate = row.get("report_date_raw")
        fdate = row.get("filing_date_raw")
        if not _dates_match_target_years(row.get("report_date"), row.get("filing_date"), target_years):
            continue
        idx = _load_ex99_accession_index(sec, cik_int, accn_nd, ex99_runtime_cache)
        if idx is None:
            continue
        exdocs = list(row.get("exdocs") or [])
        if not exdocs:
            imgs = _load_ex99_index_images(sec, cik_int, accn_nd, idx, ex99_runtime_cache)
            if imgs:
                try:
                    txt = sec.ocr_html_assets(
                        accn_nd,
                        None,
                        context={"doc": "index_images", "purpose": "eps_ocr", "report_date": rdate, "filing_date": fdate, "save_text": True},
                    )
                except Exception:
                    txt = ""
                q_end = infer_quarter_end_from_text(txt) or parse_date(rdate) or parse_date(fdate)
                if not _is_quarter_end(q_end):
                    q_end = _coerce_prev_quarter_end(q_end)
                if q_end is not None and txt and q_end not in out and _accept_q(q_end):
                    res = _extract_eps_shares_from_html(txt.encode("utf-8", errors="ignore"), q_end)
                    if not res:
                        res = _extract_eps_shares_from_text(txt, q_end)
                    if res:
                        res["label_diluted"] = (res.get("label_diluted") or "") + " (ocr)"
                        out[q_end] = res
                        if target_quarters is not None:
                            target_quarters.discard(q_end)
                        sec.ocr_log_rows.append({
                            "accn": accn,
                            "doc": "index_images",
                            "quarter": q_end,
                            "purpose": "eps_ocr_match",
                            "status": "ok",
                            "ocr_tokens": res.get("label_diluted"),
                        })
                        audit_rows.append({
                            "quarter": q_end,
                            "source": "tier3_ex99_eps",
                            "accn": accn,
                            "doc": "index_images",
                            "note": "EPS shares from image-only 8-K (OCR)",
                        })
            continue
        for fn in exdocs[:6]:
            b = _load_ex99_document_bytes(sec, cik_int, accn_nd, fn, ex99_runtime_cache)
            if b is None:
                continue
            is_pdf = fn.lower().endswith(".pdf")
            try:
                sec.download_html_assets(cik_int, accn_nd, b)
            except Exception:
                pass
            _load_ex99_index_images(sec, cik_int, accn_nd, idx, ex99_runtime_cache)
            raw = ""
            txt = ""
            if not is_pdf:
                raw = b.decode("utf-8", errors="ignore")
                txt = strip_html(raw)
            else:
                txt = _extract_text_from_pdf_bytes(b, quiet_pdf_warnings=quiet_pdf_warnings)
            q_end = infer_quarter_end_from_text(txt)
            if not _is_quarter_end(q_end):
                q_end = _coerce_prev_quarter_end(parse_date(rdate) or parse_date(fdate))
            # Prefer table-based extraction; fallback to raw row parse for EX-99 HTML
            res = _extract_eps_shares_from_html(b, q_end) if (q_end and not is_pdf) else None
            if not res:
                if not is_pdf:
                    m = re.search(r"weighted[- ]?average shares.*diluted", raw, re.I)
                    if m:
                        row_txt = raw[m.start():m.start() + 2000]
                        nums = re.findall(r"(-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d{4,}(?:\.\d+)?)", row_txt)
                        if nums:
                            val = float(nums[0].replace(",", ""))
                            if re.search(r"in thousands", raw, re.I):
                                val *= 1000.0
                            if val < 5_000_000:
                                val *= 1000.0
                            if not (0 < val <= 5_000_000_000):
                                continue
                            res = {
                                "shares_diluted": val,
                                "shares_basic": None,
                                "label_diluted": "weighted-average shares used in diluted earnings per share (ex99 html)",
                                "label_basic": "",
                            }
                elif txt:
                    res = _extract_eps_shares_from_text(txt, q_end)
            if not res:
                try:
                    ocr_txt = sec.ocr_html_assets(
                        accn_nd,
                        b,
                        context={
                            "doc": fn,
                            "quarter": q_end,
                            "purpose": "eps_ocr",
                            "report_date": rdate,
                            "filing_date": fdate,
                            "save_text": True,
                        },
                    )
                except Exception:
                    ocr_txt = ""
                if ocr_txt:
                    if not q_end:
                        q_end = infer_quarter_end_from_text(ocr_txt)
                    if not _is_quarter_end(q_end):
                        q_end = _coerce_prev_quarter_end(q_end or parse_date(rdate) or parse_date(fdate))
                if q_end is None:
                    continue
                res = _extract_eps_shares_from_html(ocr_txt.encode("utf-8", errors="ignore"), q_end)
                if not res:
                    res = _extract_eps_shares_from_text(ocr_txt, q_end)
                    if res:
                        res["label_diluted"] = (res.get("label_diluted") or "") + " (ocr)"
            if not res:
                continue
            if res.get("label_diluted") and "ocr" in res.get("label_diluted", "").lower():
                sec.ocr_log_rows.append({
                    "accn": accn,
                    "doc": fn,
                    "quarter": q_end,
                    "purpose": "eps_ocr_match",
                    "status": "ok",
                    "ocr_tokens": res.get("label_diluted"),
                })
            if q_end is None or q_end in out or not _accept_q(q_end):
                continue
            out[q_end] = res
            if target_quarters is not None:
                target_quarters.discard(q_end)
            audit_rows.append({
                "quarter": q_end,
                "source": "tier3_ex99_eps",
                "accn": accn,
                "doc": fn,
                "note": "EPS shares from EX-99.1",
            })
            break

    # Fallback: scan cached EX-99 files (including OCR) for older years
    scanned = 0
    for entry in ex99_inventory.get("legacy_docs", []) or []:
        if target_quarters is not None and not target_quarters:
            break
        if scanned >= 220:
            break
        path_in = entry.get("path")
        if not isinstance(path_in, Path):
            continue
        accn_nd = entry.get("accn_nd")
        fn = entry.get("doc") or entry.get("name")
        try:
            if entry.get("is_pdf"):
                raw = ""
            else:
                raw = path_in.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        txt = strip_html(raw) if raw else ""
        q_end = infer_quarter_end_from_text(txt)
        if not _is_quarter_end(q_end) and accn_nd:
            rep, fdate = accn_dates.get(accn_nd, (None, None))
            if target_years is not None and not _dates_match_target_years(rep, fdate, target_years):
                continue
            q_end = rep or fdate
        if q_end is not None and not _is_quarter_end(q_end):
            q_end = _coerce_prev_quarter_end(q_end)
        m = re.search(r"weighted[- ]?average shares.*diluted", raw, re.I) if raw else None
        if not m:
            # Try OCR on cached EX-99 assets (image-based statements)
            if not accn_nd:
                continue
            b = _load_legacy_ex99_document_bytes(path_in, ex99_runtime_cache)
            if b is None:
                continue
            try:
                sec.download_html_assets(cik_int, accn_nd, b)
            except Exception:
                pass
            if entry.get("is_pdf"):
                txt = _extract_text_from_pdf_bytes(b, quiet_pdf_warnings=quiet_pdf_warnings)
                q_end = infer_quarter_end_from_text(txt)
                if not _is_quarter_end(q_end):
                    rep, fdate = accn_dates.get(accn_nd, (None, None))
                    q_end = rep or fdate
                if q_end is not None and not _is_quarter_end(q_end):
                    q_end = _coerce_prev_quarter_end(q_end)
                if _is_quarter_end(q_end) and q_end not in out:
                    res = _extract_eps_shares_from_text(txt, q_end)
                    if res:
                        res["label_diluted"] = (res.get("label_diluted") or "") + " (pdf)"
                        out[q_end] = res
                        audit_rows.append({
                            "quarter": q_end,
                            "source": "tier3_ex99_eps",
                            "accn": accn_nd,
                            "doc": fn,
                            "note": "EPS shares from cached EX-99 (pdf text)",
                        })
                        scanned += 1
                        continue
            try:
                ocr_txt = sec.ocr_html_assets(
                    accn_nd,
                    b,
                    context={"doc": fn, "purpose": "eps_ocr_cache", "save_text": False},
                )
            except Exception:
                ocr_txt = ""
            scanned += 1
            if not ocr_txt:
                continue
            q_end = infer_quarter_end_from_text(ocr_txt)
            if not _is_quarter_end(q_end):
                rep, fdate = accn_dates.get(accn_nd, (None, None))
                q_end = _coerce_prev_quarter_end(q_end or rep or fdate)
            if not _is_quarter_end(q_end) or q_end in out:
                continue
            res = _extract_eps_shares_from_html(ocr_txt.encode("utf-8", errors="ignore"), q_end)
            if not res:
                res = _extract_eps_shares_from_text(ocr_txt, q_end)
            if not res:
                continue
            res["label_diluted"] = (res.get("label_diluted") or "") + " (ocr)"
            out[q_end] = res
            sec.ocr_log_rows.append({
                "accn": accn_nd,
                "doc": fn,
                "quarter": q_end,
                "purpose": "eps_ocr_match",
                "status": "ok",
                "ocr_tokens": res.get("label_diluted"),
            })
            audit_rows.append({
                "quarter": q_end,
                "source": "tier3_ex99_eps",
                "accn": accn_nd,
                "doc": fn,
                "note": "EPS shares from cached EX-99 (OCR)",
            })
            continue
        if q_end is None or not _is_quarter_end(q_end) or q_end in out or not _accept_q(q_end):
            continue
        row_txt = raw[m.start():m.start() + 2000]
        nums = re.findall(r"(-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d{4,}(?:\.\d+)?)", row_txt)
        if not nums:
            continue
        val = float(nums[0].replace(",", ""))
        if re.search(r"in\s+thousands", raw, re.I):
            val *= 1000.0
        if val < 5_000_000:
            val *= 1000.0
        if not (0 < val <= 5_000_000_000):
            continue
        out[q_end] = {
            "shares_diluted": val,
            "shares_basic": None,
            "label_diluted": "weighted-average shares used in diluted earnings per share (ex99 cache)",
            "label_basic": "",
        }
        if target_quarters is not None:
            target_quarters.discard(q_end)
        scanned += 1
    if out:
        qs = sorted(out.keys())[-max_quarters:]
        out = {q: out[q] for q in qs}
    return out, audit_rows




def build_gaap_history(
    df_all: pd.DataFrame,
    max_quarters: int = 80,
    strictness: str = "ytd",
    min_year: Optional[int] = None,
    *,
    sec: SecClient | None = None,
    cik_int: int | None = None,
    submissions: Dict[str, Any] | None = None,
    ticker: Optional[str] = None,
    quiet_pdf_warnings: bool = True,
    stage_timings: Optional[Dict[str, float]] = None,
    profile_timings: bool = False,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    gaap_stage_timings = stage_timings if stage_timings is not None else {}
    ends, fy_fp_to_end = build_quarter_calendar_from_revenue(df_all, max_quarters=max_quarters)
    # Enrich quarter calendar/mapping beyond the single "best revenue tag" so year-end
    # 10-K data (FY/Q4) is still included when tags switch between filings.
    extra_ends: set[dt.date] = set(ends)
    if df_all is not None and not df_all.empty and "end_d" in df_all.columns:
        cal = df_all[df_all["end_d"].notna()].copy()
        if "form" in cal.columns:
            forms = cal["form"].astype(str).str.upper()
            cal = cal[forms.str.startswith(("10-Q", "10-K"))]
        if "fp" in cal.columns:
            fp_u = cal["fp"].astype(str).str.upper().str.strip()
            fy_like = cal[fp_u.isin(["FY", "Q4"])]
            for d in fy_like["end_d"].dropna():
                d0 = pd.to_datetime(d, errors="coerce")
                if pd.notna(d0):
                    dd = d0.date()
                    if _is_quarter_end(dd):
                        extra_ends.add(dd)

            fy_anchor = (12, 31)
            try:
                fy_rows = cal[cal["fp"].astype(str).str.upper().isin(["FY", "Q4"]) & cal["end_d"].notna()].copy()
                if not fy_rows.empty:
                    mmdd = fy_rows["end_d"].map(lambda d: (d.month, d.day))
                    if not mmdd.empty:
                        fy_anchor = mmdd.value_counts().idxmax()
            except Exception:
                fy_anchor = (12, 31)

            def _fy_from_row_end(end_ts: pd.Timestamp) -> int:
                return int(end_ts.year) + (
                    1 if (int(end_ts.month), int(end_ts.day)) > (int(fy_anchor[0]), int(fy_anchor[1])) else 0
                )

            # Build fiscal-year-end map so fp->end mappings can be validated.
            fy_end_by_year: Dict[int, dt.date] = {}
            for _, r in cal.iterrows():
                fp_raw = str(r.get("fp") or "").upper().strip()
                if fp_raw not in {"FY", "Q4"}:
                    continue
                end_d = pd.to_datetime(r.get("end_d"), errors="coerce")
                if pd.isna(end_d):
                    continue
                fy_raw = pd.to_numeric(r.get("fy_calc"), errors="coerce")
                fy_v = int(fy_raw) if pd.notna(fy_raw) else _fy_from_row_end(end_d)
                filed_d = pd.to_datetime(r.get("filed_d"), errors="coerce")
                prev = fy_end_by_year.get(fy_v)
                if prev is None:
                    fy_end_by_year[fy_v] = end_d.date()
                    continue
                # Prefer fresher FY-end mapping by filing date when available.
                prev_rows = cal[
                    cal["end_d"].notna()
                    & (cal["end_d"] == prev)
                    & cal["fp"].astype(str).str.upper().isin(["FY", "Q4"])
                ]
                prev_filed = pd.to_datetime(prev_rows["filed_d"], errors="coerce").max() if not prev_rows.empty else pd.NaT
                if pd.notna(filed_d) and (pd.isna(prev_filed) or filed_d >= prev_filed):
                    fy_end_by_year[fy_v] = end_d.date()

            # Extend (fy, fp)->end mapping with the freshest filing per key.
            map_rank: Dict[Tuple[int, str], pd.Timestamp] = {}
            for k in list(fy_fp_to_end.keys()):
                map_rank[k] = pd.Timestamp("1900-01-01")
            for _, r in cal.iterrows():
                fp_raw = str(r.get("fp") or "").upper().strip()
                if fp_raw not in {"Q1", "Q2", "Q3", "Q4", "FY"}:
                    continue
                fp_norm = "FY" if fp_raw == "Q4" else fp_raw
                end_d = pd.to_datetime(r.get("end_d"), errors="coerce")
                if pd.isna(end_d):
                    continue
                fy_raw = pd.to_numeric(r.get("fy_calc"), errors="coerce")
                if pd.isna(fy_raw):
                    fy_v = _fy_from_row_end(end_d)
                else:
                    fy_v = int(fy_raw)
                if fp_norm in {"Q1", "Q2", "Q3"}:
                    fy_end = fy_end_by_year.get(fy_v) or fy_fp_to_end.get((fy_v, "FY"))
                    if fy_end is None:
                        months_fwd = {"Q1": 9, "Q2": 6, "Q3": 3}.get(fp_norm, 0)
                        fy_end = (end_d + pd.DateOffset(months=months_fwd) + pd.offsets.MonthEnd(0)).date()
                    exp = quarter_ends_for_fy(fy_end).get(fp_norm)
                    if exp is not None and exp != end_d.date():
                        # Comparative rows in later filings can carry mismatched fp labels;
                        # keep deterministic quarter mapping only when end-date matches fp.
                        continue
                key = (fy_v, fp_norm)
                filed_d = pd.to_datetime(r.get("filed_d"), errors="coerce")
                prev_rank = map_rank.get(key, pd.Timestamp("1900-01-01"))
                if key not in fy_fp_to_end or (pd.notna(filed_d) and filed_d >= prev_rank):
                    fy_fp_to_end[key] = end_d.date()
                    map_rank[key] = filed_d if pd.notna(filed_d) else prev_rank

    # Also include latest 10-K reportDate anchors from submissions when available.
    if submissions is not None:
        rec = ((submissions or {}).get("filings") or {}).get("recent") or {}
        forms = list(rec.get("form") or [])
        reports = list(rec.get("reportDate") or [])
        filing_dates = list(rec.get("filingDate") or [])
        n = min(len(forms), len(reports))
        for i in range(n):
            form = str(forms[i] or "").upper().strip()
            if not form.startswith("10-K"):
                continue
            rep_d = parse_date(reports[i]) if i < len(reports) else None
            rep_d = rep_d if _is_quarter_end(rep_d) else _coerce_prev_quarter_end(rep_d)
            if rep_d is None:
                fil_d = parse_date(filing_dates[i]) if i < len(filing_dates) else None
                rep_d = _coerce_prev_quarter_end(fil_d)
            if rep_d is None:
                continue
            extra_ends.add(rep_d)
            fy_fp_to_end[(rep_d.year, "FY")] = rep_d

    ends = sorted(extra_ends)
    if max_quarters and len(ends) > max_quarters:
        ends = ends[-max_quarters:]
    if min_year is not None:
        ends = [d for d in ends if d and d.year >= min_year]

    audit_rows: List[Dict[str, Any]] = []
    qfd_preview_rows: List[Dict[str, Any]] = []
    qfd_unused_rows: List[Dict[str, Any]] = []
    hist = pd.DataFrame({"quarter": ends})
    allow_ytd_when_missing_3m = {
        "cfo",
        "capex",
        "da",
        "interest_paid",
        "tax_paid",
        "buybacks_cash",
        "dividends_cash",
        "acquisitions_cash",
        "debt_repayment",
        "debt_issuance",
        "research_and_development",
    }

    def quarter_index(end: dt.date) -> Optional[int]:
        # Primary anchor: calendar quarter-end date itself.
        # SEC comparative facts often carry mixed fp labels for the same end-date.
        md = (end.month, end.day)
        if md == (3, 31):
            return 1
        if md == (6, 30):
            return 2
        if md == (9, 30):
            return 3
        if md == (12, 31):
            return 4

        sub = df_all[df_all["end_d"] == end].copy()
        fps: List[str] = []
        if not sub.empty and "fp" in sub.columns:
            fps = [str(x).upper().strip() for x in sub["fp"].dropna().tolist() if str(x).strip()]
        # Prefer explicit quarter tags over FY comparatives that can appear in later filings.
        if "Q1" in fps:
            return 1
        if "Q2" in fps:
            return 2
        if "Q3" in fps:
            return 3
        if "Q4" in fps:
            return 4
        if "FY" in fps:
            return 4
        return None

    def _prep_facts(tags: List[str]) -> Tuple[Optional[str], pd.DataFrame]:
        cand = df_all[df_all["tag"].isin(tags)].copy()
        if cand.empty:
            return None, cand
        spec = MetricSpec("tmp", tags, "duration", "USD", ["10-Q", "10-K"])
        cand = _filter_unit(cand, spec)
        best_tag = choose_best_tag(cand, spec)
        if not best_tag:
            return None, cand
        return best_tag, cand[cand["tag"] == best_tag].copy()

    def _pick_duration_from_facts(
        facts: pd.DataFrame,
        end: dt.date,
        qi: int,
        prefer_forms: List[str],
        *,
        allow_ytd: bool = True,
        allow_negative: bool = True,
        allow_override: bool = True,
        max_filed_gap_days: int = 200,
        metric_name: Optional[str] = None,
    ) -> Optional[PickResult]:
        if facts is None or facts.empty:
            return None
        rec = pick_best_duration(facts, end=end, target="3M", prefer_forms=prefer_forms)
        if rec is not None:
            return PickResult(
                value=float(rec["val"]),
                source="direct",
                tag=str(rec["tag"]),
                accn=str(rec["accn"]),
                form=str(rec["form"]),
                filed=rec["filed_d"],
                start=rec["start_d"],
                end=rec["end_d"],
                unit=str(rec["unit"]),
                duration_days=int((rec["end_d"] - rec["start_d"]).days),
                note="picked direct 3M",
            )
        metric_key = str(metric_name or "").strip().lower()
        ytd_allowed_when_only3m = metric_key in allow_ytd_when_missing_3m
        try_ytd = bool(
            allow_ytd
            and qi
            and (
                strictness != "only3m"
                or ytd_allowed_when_only3m
            )
        )
        if not try_ytd:
            return None
        pr = derive_quarter_from_ytd(
            facts=facts,
            end=end,
            quarter_index=qi,
            fy_fp_to_end=fy_fp_to_end,
            prefer_forms=prefer_forms,
            allow_negative=allow_negative,
            allow_override=allow_override,
            max_filed_gap_days=max_filed_gap_days,
        )
        if pr is not None and pr.value is not None:
            note_prefix = "derived_ytd attempted after no 3M fact"
            pr.note = f"{note_prefix}; {pr.note}" if pr.note else note_prefix
            return pr
        return None

    def _pick_duration_from_tags(
        tags: List[str],
        facts: pd.DataFrame,
        end: dt.date,
        qi: int,
        prefer_forms: List[str],
        *,
        allow_ytd: bool = True,
        allow_negative: bool = True,
        allow_override: bool = True,
        max_filed_gap_days: int = 200,
        metric_name: Optional[str] = None,
    ) -> Optional[PickResult]:
        if facts is None or facts.empty:
            return None
        for tag in tags:
            sub = facts[facts["tag"] == tag]
            if sub.empty:
                continue
            pr = _pick_duration_from_facts(
                sub,
                end,
                qi,
                prefer_forms,
                allow_ytd=allow_ytd,
                allow_negative=allow_negative,
                allow_override=allow_override,
                max_filed_gap_days=max_filed_gap_days,
                metric_name=metric_name,
            )
            if pr is not None and pr.value is not None:
                return pr
        return None

    opex_tag, opex_facts = _prep_facts(["OperatingExpenses"])
    sga_tag, sga_facts = _prep_facts(["SellingGeneralAndAdministrativeExpense"])
    rd_tag, rd_facts = _prep_facts(["ResearchAndDevelopmentExpense"])
    da_primary_tag, da_primary_facts = _prep_facts(["DepreciationAndAmortization"])
    da_alt_tag, da_alt_facts = _prep_facts([
        "DepreciationDepletionAndAmortization",
        "DepreciationDepletionAndAmortizationExpense",
        "DepreciationAmortizationAndAccretionNet",
    ])
    dep_tag, dep_facts = _prep_facts(["Depreciation"])
    amort_tag, amort_facts = _prep_facts(["AmortizationOfIntangibleAssets"])
    pretax_tags = [
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxes",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterest",
    ]
    pretax_facts = df_all[df_all["tag"].isin(pretax_tags)].copy()
    if not pretax_facts.empty:
        pretax_facts = _filter_unit(pretax_facts, MetricSpec("pretax", pretax_tags, "duration", "USD", ["10-Q", "10-K"]))

    interest_exp_tags = [
        "InterestExpense",
        "InterestIncomeExpenseNet",
        "InterestExpenseNet",
        "InterestExpenseNetOfCapitalizedInterest",
        "InterestExpenseNetOfCapitalizedInterestAndDividendIncome",
    ]
    interest_exp_facts = df_all[df_all["tag"].isin(interest_exp_tags)].copy()
    if not interest_exp_facts.empty:
        interest_exp_facts = _filter_unit(interest_exp_facts, MetricSpec("interest", interest_exp_tags, "duration", "USD", ["10-Q", "10-K"]))

    tax_exp_tags = [
        "IncomeTaxExpenseBenefit",
        "IncomeTaxExpenseBenefitContinuingOperations",
    ]
    tax_exp_facts = df_all[df_all["tag"].isin(tax_exp_tags)].copy()
    if not tax_exp_facts.empty:
        tax_exp_facts = _filter_unit(tax_exp_facts, MetricSpec("tax", tax_exp_tags, "duration", "USD", ["10-Q", "10-K"]))
    cogs_comp_tags = [
        "CostOfServicesLicensesAndMaintenanceAgreements",
        "CostOfServicesMaintenanceCosts",
        "OtherCostOfServices",
    ]
    cogs_comp_facts: Dict[str, pd.DataFrame] = {}
    for _t in cogs_comp_tags:
        _tag, _facts = _prep_facts([_t])
        if _tag:
            cogs_comp_facts[_tag] = _facts

    derived_formula_keys: set[Tuple[str, str]] = set()

    for spec in GAAP_SPECS:
        spec_tags = list(spec.tags or [])
        if str(ticker or "").strip().upper() == "GPRE" and spec.name == "debt_issuance":
            spec_tags = [t for t in spec_tags if t != "ProceedsFromShortTermDebt"]
        spec_for_pick = spec if tuple(spec_tags) == tuple(spec.tags or []) else MetricSpec(
            spec.name,
            spec_tags,
            spec.kind,
            spec.unit,
            spec.prefer_forms,
        )

        if spec.name == "da":
            col = []
            for end in ends:
                qi = quarter_index(end) or 0
                pr_da = _pick_duration_from_facts(
                    da_primary_facts,
                    end,
                    qi,
                    spec.prefer_forms,
                    metric_name="da",
                ) if da_primary_tag else None
                if pr_da is None and da_alt_tag:
                    pr_da = _pick_duration_from_facts(
                        da_alt_facts,
                        end,
                        qi,
                        spec.prefer_forms,
                        metric_name="da",
                    )
                if pr_da is not None and pr_da.value is not None:
                    col.append(float(pr_da.value))
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": pr_da.source,
                        "tag": pr_da.tag,
                        "accn": pr_da.accn,
                        "form": pr_da.form,
                        "filed": pr_da.filed,
                        "start": pr_da.start,
                        "end": pr_da.end,
                        "unit": pr_da.unit,
                        "duration_days": pr_da.duration_days,
                        "value": pr_da.value,
                        "note": "D&A from combined tag",
                    })
                    continue

                pr_dep = _pick_duration_from_facts(
                    dep_facts,
                    end,
                    qi,
                    spec.prefer_forms,
                    metric_name="da",
                ) if dep_tag else None
                pr_am = _pick_duration_from_facts(
                    amort_facts,
                    end,
                    qi,
                    spec.prefer_forms,
                    metric_name="da",
                ) if amort_tag else None
                if pr_dep is not None and pr_dep.value is not None and pr_am is not None and pr_am.value is not None:
                    val = float(pr_dep.value) + float(pr_am.value)
                    src = "derived_formula"
                    if pr_dep.source in ("derived_ytd", "derived_ytd_q4") or pr_am.source in ("derived_ytd", "derived_ytd_q4"):
                        src = "derived_ytd"
                    if pr_dep.source == "derived_ytd_q4" or pr_am.source == "derived_ytd_q4":
                        src = "derived_ytd_q4"
                    col.append(val)
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": src,
                        "tag": f"{pr_dep.tag}+{pr_am.tag}",
                        "accn": pr_dep.accn or pr_am.accn,
                        "form": pr_dep.form or pr_am.form,
                        "filed": pr_dep.filed or pr_am.filed,
                        "start": pr_dep.start,
                        "end": pr_dep.end,
                        "unit": pr_dep.unit,
                        "duration_days": pr_dep.duration_days,
                        "value": val,
                        "note": f"D&A = Depreciation ({pr_dep.tag}, source={pr_dep.source}) + Amortization ({pr_am.tag}, source={pr_am.source})",
                    })
                    derived_formula_keys.add(("da", str(end)))
                    continue

                col.append(pd.NA)
                audit_rows.append({
                    "metric": spec.name,
                    "quarter": end,
                    "source": "missing",
                    "note": "no D&A components found",
                })

            hist[spec.name] = col
            continue
        if spec.name == "total_debt":
            col = []
            for end in ends:
                pr = compute_total_debt_instant(df_all, end=end, prefer_forms=spec.prefer_forms)
                if pr is None or pr.value is None:
                    col.append(pd.NA)
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": "missing",
                        "note": "no debt tags found",
                    })
                else:
                    col.append(float(pr.value))
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": pr.source,
                        "source_choice": pr.source_choice,
                        "tag": pr.tag,
                        "accn": pr.accn,
                        "form": pr.form,
                        "filed": pr.filed,
                        "start": pr.start,
                        "end": pr.end,
                        "unit": pr.unit,
                        "duration_days": pr.duration_days,
                        "value": pr.value,
                        "note": pr.note,
                    })
            hist[spec.name] = col
            continue
        if spec.name == "debt_core":
            col = []
            for end in ends:
                pr = compute_debt_core_instant(df_all, end=end, prefer_forms=spec.prefer_forms)
                if pr is None or pr.value is None:
                    col.append(pd.NA)
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": "missing",
                        "note": "no debt core tags found",
                    })
                else:
                    col.append(float(pr.value))
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": pr.source,
                        "source_choice": pr.source_choice,
                        "tag": pr.tag,
                        "accn": pr.accn,
                        "form": pr.form,
                        "filed": pr.filed,
                        "start": pr.start,
                        "end": pr.end,
                        "unit": pr.unit,
                        "duration_days": pr.duration_days,
                        "value": pr.value,
                        "note": pr.note,
                    })
            hist[spec.name] = col
            continue
        if spec.name == "lease_liabilities":
            col = []
            for end in ends:
                pr = compute_lease_liabilities_instant(df_all, end=end, prefer_forms=spec.prefer_forms)
                if pr is None or pr.value is None:
                    col.append(pd.NA)
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": "missing",
                        "note": "no lease liability tags found",
                    })
                else:
                    col.append(float(pr.value))
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": pr.source,
                        "tag": pr.tag,
                        "accn": pr.accn,
                        "form": pr.form,
                        "filed": pr.filed,
                        "start": pr.start,
                        "end": pr.end,
                        "unit": pr.unit,
                        "duration_days": pr.duration_days,
                        "value": pr.value,
                        "note": pr.note,
                    })
            hist[spec.name] = col
            continue
        if spec.name == "bank_deposits":
            col = []
            for end in ends:
                pr = compute_bank_deposits_instant(df_all, end=end, prefer_forms=spec.prefer_forms)
                if pr is None or pr.value is None:
                    col.append(pd.NA)
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": "missing",
                        "note": "no bank deposit tags found",
                    })
                else:
                    col.append(float(pr.value))
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": pr.source,
                        "tag": pr.tag,
                        "accn": pr.accn,
                        "form": pr.form,
                        "filed": pr.filed,
                        "start": pr.start,
                        "end": pr.end,
                        "unit": pr.unit,
                        "duration_days": pr.duration_days,
                        "value": pr.value,
                        "note": pr.note,
                    })
            hist[spec.name] = col
            continue
        if spec.name == "bank_finance_receivables":
            col = []
            for end in ends:
                pr = compute_bank_finance_receivables_instant(df_all, end=end, prefer_forms=spec.prefer_forms)
                if pr is None or pr.value is None:
                    col.append(pd.NA)
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": "missing",
                        "note": "no bank finance receivable tags found",
                    })
                else:
                    col.append(float(pr.value))
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": pr.source,
                        "tag": pr.tag,
                        "accn": pr.accn,
                        "form": pr.form,
                        "filed": pr.filed,
                        "start": pr.start,
                        "end": pr.end,
                        "unit": pr.unit,
                        "duration_days": pr.duration_days,
                        "value": pr.value,
                        "note": pr.note,
                    })
            hist[spec.name] = col
            continue
        if spec.name == "bank_net_funding":
            col = []
            for end in ends:
                deposits = None
                receivables = None
                if "bank_deposits" in hist.columns:
                    deposits = pd.to_numeric(hist.loc[hist["quarter"] == end, "bank_deposits"].iloc[0], errors="coerce")
                if "bank_finance_receivables" in hist.columns:
                    receivables = pd.to_numeric(hist.loc[hist["quarter"] == end, "bank_finance_receivables"].iloc[0], errors="coerce")
                if deposits is None or pd.isna(deposits) or receivables is None or pd.isna(receivables):
                    col.append(pd.NA)
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": "missing",
                        "note": "bank_net_funding requires deposits and finance receivables",
                    })
                else:
                    val = float(deposits) - float(receivables)
                    col.append(val)
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": "derived_formula",
                        "tag": "bank_deposits - bank_finance_receivables",
                        "accn": None,
                        "form": None,
                        "filed": None,
                        "start": None,
                        "end": end,
                        "unit": "USD",
                        "duration_days": None,
                        "value": val,
                        "note": "bank_net_funding = bank_deposits - bank_finance_receivables",
                    })
            hist[spec.name] = col
            continue
        if spec.name == "pension_obligation_net":
            col = []
            for end in ends:
                # Prefer explicit combined pension+OPEB liabilities when available.
                rec_cur = _pick_instant_tag(
                    df_all,
                    end=end,
                    tag="PensionAndOtherPostretirementDefinedBenefitPlansCurrentLiabilities",
                    prefer_forms=spec.prefer_forms,
                )
                rec_non = _pick_instant_tag(
                    df_all,
                    end=end,
                    tag="PensionAndOtherPostretirementDefinedBenefitPlansLiabilitiesNoncurrent",
                    prefer_forms=spec.prefer_forms,
                )

                rec_parts: List[pd.Series] = [x for x in [rec_cur, rec_non] if x is not None and pd.notna(x.get("val"))]
                if rec_parts:
                    val = float(sum(float(r.get("val")) for r in rec_parts))
                    last_rec = sorted(
                        rec_parts,
                        key=lambda r: pd.to_datetime(r.get("filed_d"), errors="coerce")
                        if pd.notna(pd.to_datetime(r.get("filed_d"), errors="coerce"))
                        else pd.Timestamp("1900-01-01"),
                    )[-1]
                    tags_used = ",".join([str(r.get("tag")) for r in rec_parts])
                    col.append(val)
                    audit_rows.append(
                        {
                            "metric": spec.name,
                            "quarter": end,
                            "source": "derived_parts" if len(rec_parts) > 1 else "direct",
                            "tag": tags_used,
                            "accn": str(last_rec.get("accn")) if last_rec is not None else None,
                            "form": str(last_rec.get("form")) if last_rec is not None else None,
                            "filed": last_rec.get("filed_d") if last_rec is not None else None,
                            "start": last_rec.get("start_d") if last_rec is not None else None,
                            "end": last_rec.get("end_d") if last_rec is not None else None,
                            "unit": str(last_rec.get("unit")) if last_rec is not None else "USD",
                            "duration_days": None,
                            "value": val,
                            "note": "pension_obligation_net from current + noncurrent defined-benefit liabilities",
                        }
                    )
                    continue

                # Fallback: sum separate pension and other-postretirement noncurrent liabilities.
                rec_pen_non = _pick_instant_tag(
                    df_all,
                    end=end,
                    tag="DefinedBenefitPensionPlanLiabilitiesNoncurrent",
                    prefer_forms=spec.prefer_forms,
                )
                rec_opeb_non = _pick_instant_tag(
                    df_all,
                    end=end,
                    tag="OtherPostretirementDefinedBenefitPlanLiabilitiesNoncurrent",
                    prefer_forms=spec.prefer_forms,
                )
                rec_cur_alt = _pick_instant_tag(
                    df_all,
                    end=end,
                    tag="PensionAndOtherPostretirementDefinedBenefitPlansCurrentLiabilities",
                    prefer_forms=spec.prefer_forms,
                )
                rec_parts_alt: List[pd.Series] = [
                    x
                    for x in [rec_pen_non, rec_opeb_non, rec_cur_alt]
                    if x is not None and pd.notna(x.get("val"))
                ]
                if rec_parts_alt:
                    val = float(sum(float(r.get("val")) for r in rec_parts_alt))
                    last_rec = sorted(
                        rec_parts_alt,
                        key=lambda r: pd.to_datetime(r.get("filed_d"), errors="coerce")
                        if pd.notna(pd.to_datetime(r.get("filed_d"), errors="coerce"))
                        else pd.Timestamp("1900-01-01"),
                    )[-1]
                    tags_used = ",".join([str(r.get("tag")) for r in rec_parts_alt])
                    col.append(val)
                    audit_rows.append(
                        {
                            "metric": spec.name,
                            "quarter": end,
                            "source": "derived_parts",
                            "tag": tags_used,
                            "accn": str(last_rec.get("accn")) if last_rec is not None else None,
                            "form": str(last_rec.get("form")) if last_rec is not None else None,
                            "filed": last_rec.get("filed_d") if last_rec is not None else None,
                            "start": last_rec.get("start_d") if last_rec is not None else None,
                            "end": last_rec.get("end_d") if last_rec is not None else None,
                            "unit": str(last_rec.get("unit")) if last_rec is not None else "USD",
                            "duration_days": None,
                            "value": val,
                            "note": "pension_obligation_net from pension + OPEB liability components",
                        }
                    )
                    continue

                # Last resort: use direct total tags if they exist.
                rec_direct = _pick_first_instant_tag(
                    df_all,
                    end=end,
                    tags=[
                        "PensionAndOtherPostretirementObligations",
                        "PensionAndOtherPostretirementBenefitObligations",
                        "DefinedBenefitPlanAmountsRecognizedInBalanceSheet",
                    ],
                    prefer_forms=spec.prefer_forms,
                )
                if rec_direct is not None and pd.notna(rec_direct.get("val")):
                    val = float(rec_direct.get("val"))
                    note = "picked instant pension/OPEB tag"
                    # This tag is often signed opposite to liability presentation; use magnitude.
                    if str(rec_direct.get("tag")) == "DefinedBenefitPlanAmountsRecognizedInBalanceSheet":
                        val = abs(val)
                        note = "used absolute value of DefinedBenefitPlanAmountsRecognizedInBalanceSheet"
                    col.append(val)
                    audit_rows.append(
                        {
                            "metric": spec.name,
                            "quarter": end,
                            "source": "direct",
                            "tag": str(rec_direct.get("tag")),
                            "accn": str(rec_direct.get("accn")),
                            "form": str(rec_direct.get("form")),
                            "filed": rec_direct.get("filed_d"),
                            "start": rec_direct.get("start_d"),
                            "end": rec_direct.get("end_d"),
                            "unit": str(rec_direct.get("unit")),
                            "duration_days": None,
                            "value": val,
                            "note": note,
                        }
                    )
                else:
                    col.append(pd.NA)
                    audit_rows.append(
                        {
                            "metric": spec.name,
                            "quarter": end,
                            "source": "missing",
                            "note": "no pension/OPEB liability tags found",
                        }
                    )
            hist[spec.name] = col
            continue

        if spec.kind == "instant":
            col = []
            for end in ends:
                cand = df_all[df_all["tag"].isin(spec_tags)].copy()
                cand = _filter_unit(cand, spec_for_pick)
                best_tag = choose_best_tag(cand, spec_for_pick)
                if not best_tag:
                    col.append(pd.NA)
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": "missing",
                        "note": "no instant tag found",
                    })
                    continue
                facts = cand[cand["tag"] == best_tag].copy()
                rec = pick_best_instant(facts, end=end, prefer_forms=spec.prefer_forms)
                if rec is None:
                    col.append(pd.NA)
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": "missing",
                        "tag": best_tag,
                        "note": "no instant fact",
                    })
                else:
                    col.append(float(rec["val"]))
                    audit_rows.append({
                        "metric": spec.name,
                        "quarter": end,
                        "source": "direct",
                        "tag": str(rec["tag"]),
                        "accn": str(rec["accn"]),
                        "form": str(rec["form"]),
                        "filed": rec["filed_d"],
                        "start": rec["start_d"],
                        "end": rec["end_d"],
                        "unit": str(rec["unit"]),
                        "duration_days": None,
                        "value": float(rec["val"]),
                        "note": "picked instant",
                    })
            hist[spec.name] = col
            continue

        cand = df_all[df_all["tag"].isin(spec_tags)].copy()
        cand = _filter_unit(cand, spec_for_pick)
        best_tag = choose_best_tag(cand, spec_for_pick)
        if not best_tag:
            hist[spec.name] = [pd.NA] * len(ends)
            for end in ends:
                audit_rows.append({
                    "metric": spec.name,
                    "quarter": end,
                    "source": "missing",
                    "note": "no tag found",
                })
            continue
        facts = cand[cand["tag"] == best_tag].copy()
        col = []
        for end in ends:
            qi = quarter_index(end) or 0
            allow_ytd = spec.name not in ("shares_diluted",)
            allow_negative = spec.name not in ("revenue", "cogs")
            allow_override = spec.name not in (
                "revenue",
                "research_and_development",
                "buybacks_cash",
                "dividends_cash",
                "acquisitions_cash",
                "debt_repayment",
                "debt_issuance",
            )
            max_gap = 200 if spec.name in ("revenue", "cogs") else 365
            pr = _pick_duration_from_facts(
                facts,
                end,
                qi,
                spec.prefer_forms,
                allow_ytd=allow_ytd,
                allow_negative=allow_negative,
                allow_override=allow_override,
                max_filed_gap_days=max_gap,
                metric_name=spec.name,
            )

            if pr is not None and pr.value is not None:
                raw_val = float(pr.value)
                out_val = raw_val
                note_txt = pr.note
                if spec.name == "interest_expense_net":
                    out_val = abs(raw_val)
                    note_txt = (note_txt + "; " if note_txt else "") + "stored as positive interest expense magnitude"
                col.append(out_val)
                audit_entry = {
                    "metric": spec.name,
                    "quarter": end,
                    "source": pr.source,
                    "tag": pr.tag,
                    "accn": pr.accn,
                    "form": pr.form,
                    "filed": pr.filed,
                    "start": pr.start,
                    "end": pr.end,
                    "unit": pr.unit,
                    "duration_days": pr.duration_days,
                    "value": out_val,
                    "note": note_txt,
                }
                if spec.name == "interest_expense_net":
                    audit_entry["raw_value"] = raw_val
                    audit_entry["normalized_value"] = out_val
                audit_rows.append(audit_entry)
            else:
                rec = pick_best_duration(facts, end=end, target="3M", prefer_forms=spec.prefer_forms)
                if rec is None:
                    # Try alternative tags for this quarter if best_tag has no 3M fact
                    alt_pr = None
                    for alt in spec_tags:
                        if alt == best_tag:
                            continue
                        alt_facts = cand[cand["tag"] == alt].copy()
                        if alt_facts.empty:
                            continue
                        alt_pr = _pick_duration_from_facts(
                            alt_facts,
                            end,
                            qi,
                            spec.prefer_forms,
                            allow_ytd=allow_ytd,
                            allow_negative=allow_negative,
                            allow_override=allow_override,
                            max_filed_gap_days=max_gap,
                            metric_name=spec.name,
                        )
                        if alt_pr is not None and alt_pr.value is not None:
                            break
                    if alt_pr is not None and alt_pr.value is not None:
                        raw_val = float(alt_pr.value)
                        out_val = raw_val
                        note_txt = f"picked alternate tag ({alt_pr.tag})"
                        if spec.name == "interest_expense_net":
                            out_val = abs(raw_val)
                            note_txt = note_txt + "; stored as positive interest expense magnitude"
                        col.append(out_val)
                        audit_entry = {
                            "metric": spec.name,
                            "quarter": end,
                            "source": alt_pr.source,
                            "tag": alt_pr.tag,
                            "accn": alt_pr.accn,
                            "form": alt_pr.form,
                            "filed": alt_pr.filed,
                            "start": alt_pr.start,
                            "end": alt_pr.end,
                            "unit": alt_pr.unit,
                            "duration_days": alt_pr.duration_days,
                            "value": out_val,
                            "note": note_txt,
                        }
                        if spec.name == "interest_expense_net":
                            audit_entry["raw_value"] = raw_val
                            audit_entry["normalized_value"] = out_val
                        audit_rows.append(audit_entry)
                    else:
                        if spec.name == "interest_expense_net":
                            exp_tags = [
                                "InterestExpense",
                                "InterestExpenseNonoperating",
                                "InterestExpenseNetOfCapitalizedInterest",
                                "InterestExpenseNetOfCapitalizedInterestAndDividendIncome",
                            ]
                            inc_tags = [
                                "InterestIncome",
                                "InvestmentIncomeInterest",
                                "InterestAndDividendIncomeOperating",
                            ]
                            exp_facts = df_all[df_all["tag"].isin(exp_tags)].copy()
                            inc_facts = df_all[df_all["tag"].isin(inc_tags)].copy()
                            if not exp_facts.empty:
                                exp_facts = _filter_unit(exp_facts, MetricSpec("interest_exp", exp_tags, "duration", "USD", ["10-Q", "10-K"]))
                            if not inc_facts.empty:
                                inc_facts = _filter_unit(inc_facts, MetricSpec("interest_inc", inc_tags, "duration", "USD", ["10-Q", "10-K"]))
                            pr_exp = _pick_duration_from_tags(
                                exp_tags,
                                exp_facts,
                                end,
                                qi,
                                spec.prefer_forms,
                                allow_ytd=allow_ytd,
                                allow_negative=True,
                                allow_override=allow_override,
                                max_filed_gap_days=max_gap,
                                metric_name=spec.name,
                            ) if not exp_facts.empty else None
                            pr_inc = _pick_duration_from_tags(
                                inc_tags,
                                inc_facts,
                                end,
                                qi,
                                spec.prefer_forms,
                                allow_ytd=allow_ytd,
                                allow_negative=True,
                                allow_override=allow_override,
                                max_filed_gap_days=max_gap,
                                metric_name=spec.name,
                            ) if not inc_facts.empty else None
                            if pr_exp is not None and pr_exp.value is not None and pr_inc is not None and pr_inc.value is not None:
                                val_comp = abs(float(pr_exp.value) - float(pr_inc.value))
                                col.append(val_comp)
                                audit_rows.append(
                                    {
                                        "metric": spec.name,
                                        "quarter": end,
                                        "source": "derived_formula",
                                        "tag": f"{pr_exp.tag}-{pr_inc.tag}",
                                        "accn": pr_exp.accn or pr_inc.accn,
                                        "form": pr_exp.form or pr_inc.form,
                                        "filed": pr_exp.filed or pr_inc.filed,
                                        "start": pr_exp.start,
                                        "end": pr_exp.end,
                                        "unit": pr_exp.unit or pr_inc.unit,
                                        "duration_days": pr_exp.duration_days,
                                        "value": val_comp,
                                        "note": "interest_expense_net derived as positive magnitude of InterestExpense - InterestIncome",
                                    }
                                )
                                continue
                            if pr_exp is not None and pr_exp.value is not None:
                                val_comp = abs(float(pr_exp.value))
                                col.append(val_comp)
                                audit_rows.append(
                                    {
                                        "metric": spec.name,
                                        "quarter": end,
                                        "source": "derived_formula",
                                        "tag": str(pr_exp.tag or "InterestExpense"),
                                        "accn": pr_exp.accn,
                                        "form": pr_exp.form,
                                        "filed": pr_exp.filed,
                                        "start": pr_exp.start,
                                        "end": pr_exp.end,
                                        "unit": pr_exp.unit,
                                        "duration_days": pr_exp.duration_days,
                                        "value": val_comp,
                                        "note": "interest_expense_net fallback from positive expense-only magnitude",
                                    }
                                )
                                continue
                        col.append(pd.NA)
                        attempted_ytd = bool(
                            allow_ytd
                            and qi
                            and (
                                strictness != "only3m"
                                or spec.name in allow_ytd_when_missing_3m
                            )
                        )
                        audit_rows.append({
                            "metric": spec.name,
                            "quarter": end,
                            "source": "missing",
                            "tag": best_tag,
                            "note": (
                                "no 3M fact; derived_ytd attempted after no 3M fact but unavailable"
                                if attempted_ytd
                                else "no 3M fact"
                            ),
                        })
                else:
                    raw_val = float(rec["val"])
                    out_val = raw_val
                    note_txt = "picked 3M fallback"
                    if spec.name == "interest_expense_net":
                        out_val = abs(raw_val)
                        note_txt = "picked 3M fallback; stored as positive interest expense magnitude"
                    col.append(out_val)
                    audit_entry = {
                        "metric": spec.name,
                        "quarter": end,
                        "source": "direct",
                        "tag": str(rec["tag"]),
                        "accn": str(rec["accn"]),
                        "form": str(rec["form"]),
                        "filed": rec["filed_d"],
                        "start": rec["start_d"],
                        "end": rec["end_d"],
                        "unit": str(rec["unit"]),
                        "duration_days": int((rec["end_d"] - rec["start_d"]).days),
                        "value": out_val,
                        "note": note_txt,
                    }
                    if spec.name == "interest_expense_net":
                        audit_entry["raw_value"] = raw_val
                        audit_entry["normalized_value"] = out_val
                    audit_rows.append(audit_entry)

        hist[spec.name] = col

    # Conservative leading backfill for early-history debt where first quarter has no instant fact.
    # This prevents a single head NA from blanking downstream leverage displays while keeping
    # interior missing values explicit.
    for _m in ("total_debt", "debt_core"):
        if _m not in hist.columns:
            continue
        _s = pd.to_numeric(hist[_m], errors="coerce")
        fv = _s.first_valid_index()
        if fv is None:
            continue
        try:
            fv_i = int(fv)
        except Exception:
            continue
        if fv_i <= 0:
            continue
        fv_val = _s.iloc[fv_i]
        if pd.isna(fv_val):
            continue
        for idx in range(0, fv_i):
            if pd.notna(pd.to_numeric(hist.at[idx, _m], errors="coerce")):
                continue
            hist.at[idx, _m] = float(fv_val)
            q = hist.at[idx, "quarter"]
            audit_rows.append(
                {
                    "metric": _m,
                    "quarter": q,
                    "source": "derived_carryforward_first_available",
                    "value": float(fv_val),
                    "note": f"{_m} backfilled from first available quarter {hist.at[fv_i, 'quarter']}",
                }
            )

    # formula-based metrics
    # COGS: if missing and component tags exist, sum components
    if "cogs" in hist.columns and cogs_comp_facts:
        for idx, row in hist.iterrows():
            q = row["quarter"]
            if pd.notna(row.get("cogs")):
                continue
            qi = quarter_index(q) or 0
            comps = []
            comp_notes = []
            sources = []
            for tag, facts in cogs_comp_facts.items():
                pr_c = _pick_duration_from_facts(facts, q, qi, ["10-Q", "10-K"]) if facts is not None else None
                if pr_c is None or pr_c.value is None:
                    comps = []
                    break
                comps.append(float(pr_c.value))
                comp_notes.append(f"{tag}:{pr_c.source}")
                sources.append(pr_c.source)
            if comps:
                val = float(sum(comps))
                hist.at[idx, "cogs"] = val
                src = "derived_formula"
                if any(s in ("derived_ytd", "derived_ytd_q4") for s in sources):
                    src = "derived_ytd"
                if any(s == "derived_ytd_q4" for s in sources):
                    src = "derived_ytd_q4"
                audit_rows.append({
                    "metric": "cogs",
                    "quarter": q,
                    "source": src,
                    "tag": "+".join(cogs_comp_facts.keys()),
                    "accn": None,
                    "form": None,
                    "filed": None,
                    "start": None,
                    "end": q,
                    "unit": "USD",
                    "duration_days": None,
                    "value": val,
                    "note": f"cogs = sum(components) [{'; '.join(comp_notes)}]",
                })
                derived_formula_keys.add(("cogs", str(q)))
    if "cogs" in hist.columns and "gross_profit" in hist.columns:
        for idx, row in hist.iterrows():
            if pd.isna(row.get("cogs")) and pd.notna(row.get("revenue")) and pd.notna(row.get("gross_profit")):
                val = float(row["revenue"]) - float(row["gross_profit"])
                hist.at[idx, "cogs"] = val
                q = row["quarter"]
                audit_rows.append({
                    "metric": "cogs",
                    "quarter": q,
                    "source": "derived_formula",
                    "value": val,
                    "note": "cogs = revenue - gross_profit",
                })
                derived_formula_keys.add(("cogs", str(q)))

    if "gross_profit" in hist.columns:
        for idx, row in hist.iterrows():
            if pd.isna(row["gross_profit"]) and pd.notna(row.get("revenue")) and pd.notna(row.get("cogs")):
                val = float(row["revenue"]) - float(row["cogs"])
                hist.at[idx, "gross_profit"] = val
                q = row["quarter"]
                audit_rows.append({
                    "metric": "gross_profit",
                    "quarter": q,
                    "source": "derived_formula",
                    "value": val,
                    "note": "gross_profit = revenue - cogs",
                })
                derived_formula_keys.add(("gross_profit", str(q)))

    if "op_income" in hist.columns and "gross_profit" in hist.columns:
        for idx, row in hist.iterrows():
            if pd.isna(row["op_income"]) and pd.notna(row.get("gross_profit")):
                q = row["quarter"]
                qi = quarter_index(q) or 0
                gp = float(row["gross_profit"])

                pr_opex = _pick_duration_from_facts(opex_facts, q, qi, ["10-Q", "10-K"]) if opex_tag else None
                if pr_opex is not None and pr_opex.value is not None:
                    val = gp - float(pr_opex.value)
                    src = "derived_formula"
                    if pr_opex.source in ("derived_ytd", "derived_ytd_q4"):
                        src = pr_opex.source
                    hist.at[idx, "op_income"] = val
                    audit_rows.append({
                        "metric": "op_income",
                        "quarter": q,
                        "source": src,
                        "value": val,
                        "note": f"op_income = gross_profit - OperatingExpenses ({pr_opex.tag}, source={pr_opex.source})",
                    })
                    derived_formula_keys.add(("op_income", str(q)))
                    continue

                pr_sga = _pick_duration_from_facts(sga_facts, q, qi, ["10-Q", "10-K"]) if sga_tag else None
                pr_rd = _pick_duration_from_facts(
                    rd_facts,
                    q,
                    qi,
                    ["10-Q", "10-K"],
                    metric_name="research_and_development",
                ) if rd_tag else None
                if pr_sga is not None and pr_sga.value is not None and pr_rd is not None and pr_rd.value is not None:
                    val = gp - (float(pr_sga.value) + float(pr_rd.value))
                    src = "derived_formula"
                    if pr_sga.source in ("derived_ytd", "derived_ytd_q4") or pr_rd.source in ("derived_ytd", "derived_ytd_q4"):
                        src = "derived_ytd"
                    if pr_sga.source == "derived_ytd_q4" or pr_rd.source == "derived_ytd_q4":
                        src = "derived_ytd_q4"
                    hist.at[idx, "op_income"] = val
                    audit_rows.append({
                        "metric": "op_income",
                        "quarter": q,
                        "source": src,
                        "value": val,
                        "note": f"op_income = gross_profit - (SG&A {pr_sga.tag}, source={pr_sga.source} + R&D {pr_rd.tag}, source={pr_rd.source})",
                    })
                    derived_formula_keys.add(("op_income", str(q)))

    # EBITDA (GAAP proxy) = op_income + D&A
    if "op_income" in hist.columns and "da" in hist.columns:
        ebitda_vals = []
        for idx, row in hist.iterrows():
            q = row["quarter"]
            if pd.notna(row.get("op_income")) and pd.notna(row.get("da")):
                val = float(row["op_income"]) + float(row["da"])
                ebitda_vals.append(val)
                audit_rows.append({
                    "metric": "ebitda",
                    "quarter": q,
                    "source": "derived_formula",
                    "value": val,
                    "note": "ebitda = op_income + da",
                })
                derived_formula_keys.add(("ebitda", str(q)))
            else:
                ebitda_vals.append(pd.NA)
        hist["ebitda"] = ebitda_vals

    # EBITDA fallback paths (when op_income missing)
    if "ebitda" in hist.columns:
        for idx, row in hist.iterrows():
            if pd.notna(row.get("ebitda")):
                continue
            q = row["quarter"]
            qi = quarter_index(q) or 0
            da_val = row.get("da")
            if pd.isna(da_val):
                continue

            pr_pretax = _pick_duration_from_tags(pretax_tags, pretax_facts, q, qi, ["10-Q", "10-K"]) if not pretax_facts.empty else None
            pr_int = _pick_duration_from_tags(interest_exp_tags, interest_exp_facts, q, qi, ["10-Q", "10-K"]) if not interest_exp_facts.empty else None
            pr_tax = _pick_duration_from_tags(tax_exp_tags, tax_exp_facts, q, qi, ["10-Q", "10-K"]) if not tax_exp_facts.empty else None

            # Path B: pretax + interest + D&A
            if pr_pretax is not None and pr_pretax.value is not None and pr_int is not None and pr_int.value is not None:
                int_addback = float(pr_int.value)
                if pr_int.tag and "IncomeExpenseNet" in pr_int.tag:
                    int_addback = -int_addback
                val = float(pr_pretax.value) + int_addback + float(da_val)
                hist.at[idx, "ebitda"] = val
                audit_rows.append({
                    "metric": "ebitda",
                    "quarter": q,
                    "source": "derived_formula",
                    "tag": f"{pr_pretax.tag}+{pr_int.tag}+D&A",
                    "accn": pr_pretax.accn or pr_int.accn,
                    "form": pr_pretax.form or pr_int.form,
                    "filed": pr_pretax.filed or pr_int.filed,
                    "start": pr_pretax.start,
                    "end": pr_pretax.end,
                    "unit": pr_pretax.unit,
                    "duration_days": pr_pretax.duration_days,
                    "value": val,
                    "note": "ebitda = pretax + interest addback + da (path B)",
                })
                derived_formula_keys.add(("ebitda", str(q)))
                continue

            # Path C: net_income + tax_expense + interest + D&A
            net_income_val = row.get("net_income")
            if (
                pd.notna(net_income_val)
                and pr_tax is not None and pr_tax.value is not None
                and pr_int is not None and pr_int.value is not None
            ):
                int_addback = float(pr_int.value)
                if pr_int.tag and "IncomeExpenseNet" in pr_int.tag:
                    int_addback = -int_addback
                val = float(net_income_val) + float(pr_tax.value) + int_addback + float(da_val)
                hist.at[idx, "ebitda"] = val
                audit_rows.append({
                    "metric": "ebitda",
                    "quarter": q,
                    "source": "derived_formula",
                    "tag": f"NetIncome+{pr_tax.tag}+{pr_int.tag}+D&A",
                    "accn": pr_tax.accn or pr_int.accn,
                    "form": pr_tax.form or pr_int.form,
                    "filed": pr_tax.filed or pr_int.filed,
                    "start": pr_tax.start,
                    "end": pr_tax.end,
                    "unit": pr_tax.unit,
                    "duration_days": pr_tax.duration_days,
                    "value": val,
                    "note": "ebitda = net_income + tax_expense + interest addback + da (path C)",
                })
                derived_formula_keys.add(("ebitda", str(q)))

    ex99_shared_inventory: Optional[Dict[str, Any]] = None
    ex99_runtime_cache: Optional[Dict[str, Any]] = None
    ex99_inventory_target_years: Optional[set[int]] = None
    filing_shared_inventory: Optional[Dict[str, Any]] = None
    filing_runtime_cache: Optional[Dict[str, Any]] = None
    filing_inventory_target_years: Optional[set[int]] = None
    if sec is not None and submissions is not None and cik_int is not None:
        ex99_runtime_cache = _make_ex99_runtime_cache()
        ex99_inventory_target_years = {d.year for d in ends if d is not None}
        if ex99_inventory_target_years:
            ex99_inventory_target_years |= {y + 1 for y in ex99_inventory_target_years}
        filing_runtime_cache = _make_primary_filing_runtime_cache()
        filing_inventory_target_years = set(ex99_inventory_target_years or set())

    def _get_ex99_shared_inventory() -> Optional[Dict[str, Any]]:
        nonlocal ex99_shared_inventory
        if ex99_runtime_cache is None or sec is None or submissions is None or cik_int is None:
            return None
        if ex99_shared_inventory is None:
            ex99_shared_inventory = _build_ex99_accession_inventory(
                sec,
                cik_int,
                submissions,
                ex99_runtime_cache=ex99_runtime_cache,
                target_years=ex99_inventory_target_years,
            )
        return ex99_shared_inventory

    def _get_filing_shared_inventory() -> Optional[Dict[str, Any]]:
        nonlocal filing_shared_inventory
        if filing_runtime_cache is None or sec is None or submissions is None:
            return None
        if filing_shared_inventory is None:
            filing_shared_inventory = _build_primary_filing_inventory(
                sec,
                submissions,
                target_years=filing_inventory_target_years,
            )
        return filing_shared_inventory

    # EPS note fallback for diluted shares (10-Q only, strict match on quarter_end)
    if sec is not None and submissions is not None and cik_int is not None and "shares_diluted" in hist.columns:
        filing_inventory = _get_filing_shared_inventory()
        with _timed_stage(gaap_stage_timings, "gaap_history.eps_fallback", enabled=profile_timings):
            eps_map, _eps_audit = build_eps_shares_fallback(
                sec,
                cik_int,
                submissions,
                max_quarters=max_quarters,
                filing_inventory=filing_inventory,
                filing_runtime_cache=filing_runtime_cache,
            )
        if eps_map:
            for end in ends:
                if pd.notna(hist.loc[hist["quarter"] == end, "shares_diluted"]).any():
                    continue
                payload = eps_map.get(end)
                if not payload:
                    continue
                val = payload.get("shares_diluted") or payload.get("shares_basic")
                if val is None:
                    continue
                hist.loc[hist["quarter"] == end, "shares_diluted"] = float(val)
                note = payload.get("label_diluted") or payload.get("label_basic") or "eps note"
                audit_rows.append({
                    "metric": "shares_diluted",
                    "quarter": end,
                    "source": "tier2_eps_note",
                    "tag": "EPS note",
                    "accn": None,
                    "form": None,
                    "filed": None,
                    "start": None,
                    "end": end,
                    "unit": "shares",
                    "duration_days": None,
                    "value": float(val),
                    "note": f"EPS note fallback ({note})",
                })

        # EX-99 EPS fallback from 8-K if still missing
        missing_shares_ends = set(
            [d for d in ends if pd.isna(hist.loc[hist["quarter"] == d, "shares_diluted"]).any()]
        )
        ex99_map = {}
        if missing_shares_ends:
            ex99_inventory = _get_ex99_shared_inventory()
            ex99_map, _ex99_audit = build_eps_shares_fallback_ex99(
                sec,
                cik_int,
                submissions,
                max_quarters=max_quarters,
                target_quarters=missing_shares_ends,
                quiet_pdf_warnings=quiet_pdf_warnings,
                ex99_inventory=ex99_inventory,
                ex99_runtime_cache=ex99_runtime_cache,
            )
        if ex99_map:
            for end in ends:
                if pd.notna(hist.loc[hist["quarter"] == end, "shares_diluted"]).any():
                    continue
                payload = ex99_map.get(end)
                if not payload:
                    continue
                val = payload.get("shares_diluted") or payload.get("shares_basic")
                if val is None:
                    continue
                hist.loc[hist["quarter"] == end, "shares_diluted"] = float(val)
                note = payload.get("label_diluted") or payload.get("label_basic") or "ex99 eps"
                audit_rows.append({
                    "metric": "shares_diluted",
                    "quarter": end,
                    "source": "tier3_ex99_eps",
                    "tag": "EX-99.1",
                    "accn": None,
                    "form": None,
                    "filed": None,
                    "start": None,
                    "end": end,
                    "unit": "shares",
                    "duration_days": None,
                    "value": float(val),
                    "note": f"EX-99 EPS shares ({note})",
                })
        # Final fallback: scan cached EX-99 docs directly if still missing
        # (handled inside build_eps_shares_fallback_ex99 for speed/consistency)

        shares_spec = next((sp for sp in GAAP_SPECS if sp.name == "shares_diluted"), None)
        if shares_spec is not None:
            shares_facts = df_all[df_all["tag"].isin(shares_spec.tags)].copy()
            if not shares_facts.empty:
                shares_facts = _filter_unit(shares_facts, shares_spec)
            for end in ends:
                if (quarter_index(end) or 0) != 4:
                    continue
                if pd.notna(hist.loc[hist["quarter"] == end, "shares_diluted"]).any():
                    continue
                pr_fy = None
                if not shares_facts.empty:
                    pr_fy = _pick_duration_from_tags(
                        list(shares_spec.tags),
                        shares_facts,
                        end,
                        4,
                        shares_spec.prefer_forms,
                        allow_ytd=False,
                        allow_negative=False,
                        allow_override=True,
                        max_filed_gap_days=200,
                        metric_name=shares_spec.name,
                    )
                if pr_fy is not None and pr_fy.value is not None:
                    hist.loc[hist["quarter"] == end, "shares_diluted"] = float(pr_fy.value)
                    audit_rows.append(
                        {
                            "metric": "shares_diluted",
                            "quarter": end,
                            "source": pr_fy.source,
                            "source_choice": "fallback_fy",
                            "tag": pr_fy.tag,
                            "accn": pr_fy.accn,
                            "form": pr_fy.form,
                            "filed": pr_fy.filed,
                            "start": pr_fy.start,
                            "end": pr_fy.end,
                            "unit": pr_fy.unit,
                            "duration_days": pr_fy.duration_days,
                            "value": float(pr_fy.value),
                            "note": "FY-end diluted shares fallback from same-end-date FY fact",
                        }
                    )
                    continue

                prev_hist = hist[pd.to_datetime(hist["quarter"], errors="coerce") < pd.Timestamp(end)].copy()
                if prev_hist.empty:
                    continue
                prev_hist["quarter"] = pd.to_datetime(prev_hist["quarter"], errors="coerce")
                prev_hist = prev_hist.sort_values("quarter")
                prev_vals = pd.to_numeric(prev_hist.get("shares_diluted"), errors="coerce")
                prev_valid = prev_hist[prev_vals.notna()].copy()
                if prev_valid.empty:
                    continue
                carry_row = prev_valid.iloc[-1]
                carry_val = pd.to_numeric(carry_row.get("shares_diluted"), errors="coerce")
                carry_q = pd.to_datetime(carry_row.get("quarter"), errors="coerce")
                if pd.isna(carry_val) or pd.isna(carry_q):
                    continue
                hist.loc[hist["quarter"] == end, "shares_diluted"] = float(carry_val)
                audit_rows.append(
                    {
                        "metric": "shares_diluted",
                        "quarter": end,
                        "source": "carry_forward",
                        "source_choice": "carry_forward_q3",
                        "tag": "",
                        "accn": None,
                        "form": None,
                        "filed": None,
                        "start": carry_q.date(),
                        "end": end,
                        "unit": "shares",
                        "duration_days": None,
                        "value": float(carry_val),
                        "note": f"carry-forward diluted shares from prior quarter {carry_q.date().isoformat()}",
                    }
                )

    # Cover-page fallback for shares outstanding (period-end)
    if sec is not None and submissions is not None and cik_int is not None and "shares_outstanding" in hist.columns:
        filing_inventory = _get_filing_shared_inventory()
        with _timed_stage(gaap_stage_timings, "gaap_history.shares_outstanding_fallback", enabled=profile_timings):
            so_map, so_audit = build_shares_outstanding_fallback(
                sec,
                cik_int,
                submissions,
                max_quarters=max_quarters,
                filing_inventory=filing_inventory,
                filing_runtime_cache=filing_runtime_cache,
            )
        if so_map:
            for end in ends:
                if pd.notna(hist.loc[hist["quarter"] == end, "shares_outstanding"]).any():
                    continue
                payload = so_map.get(end)
                if not payload:
                    continue
                val = payload.get("shares_outstanding")
                if val is None:
                    continue
                hist.loc[hist["quarter"] == end, "shares_outstanding"] = float(val)
        if so_audit:
            audit_rows.extend(so_audit)

    # Cash taxes paid (supplemental 10-Q table) – only if XBRL missing
    if sec is not None and submissions is not None and cik_int is not None and "tax_paid" in hist.columns:
        missing_tax_ends = set(
            [d for d in ends if pd.isna(hist.loc[hist["quarter"] == d, "tax_paid"]).any()]
        )
        tax_3m = {}
        tax_6m = {}
        tax_9m = {}
        if missing_tax_ends:
            filing_inventory = _get_filing_shared_inventory()
            with _timed_stage(gaap_stage_timings, "gaap_history.cash_taxes_fallback", enabled=profile_timings):
                tax_3m, _tax_audit = build_cash_taxes_fallback_10q(
                    sec,
                    cik_int,
                    submissions,
                    max_quarters=max_quarters,
                    period_hint="3M",
                    target_quarters=missing_tax_ends,
                    filing_inventory=filing_inventory,
                    filing_runtime_cache=filing_runtime_cache,
                )
                tax_6m, _ = build_cash_taxes_fallback_10q(
                    sec,
                    cik_int,
                    submissions,
                    max_quarters=max_quarters,
                    period_hint="6M",
                    target_quarters=missing_tax_ends,
                    filing_inventory=filing_inventory,
                    filing_runtime_cache=filing_runtime_cache,
                )
                tax_9m, _ = build_cash_taxes_fallback_10q(
                    sec,
                    cik_int,
                    submissions,
                    max_quarters=max_quarters,
                    period_hint="9M",
                    target_quarters=missing_tax_ends,
                    filing_inventory=filing_inventory,
                    filing_runtime_cache=filing_runtime_cache,
                )

        if tax_3m:
            for end in ends:
                if pd.notna(hist.loc[hist["quarter"] == end, "tax_paid"]).any():
                    continue
                payload = tax_3m.get(end)
                if not payload:
                    continue
                val = payload.get("value")
                if val is None:
                    continue
                hist.loc[hist["quarter"] == end, "tax_paid"] = float(val)
                note = payload.get("label") or "cash income tax payments (supplemental)"
                audit_rows.append({
                    "metric": "tax_paid",
                    "quarter": end,
                    "source": "tier3_cash_taxes",
                    "tag": "CashIncomeTaxPaymentsNet",
                    "accn": payload.get("accn"),
                    "form": payload.get("form"),
                    "filed": parse_date(payload.get("filing_date")),
                    "start": None,
                    "end": end,
                    "unit": "USD",
                    "duration_days": None,
                    "value": float(val),
                    "note": f"Cash taxes from 10-Q supplemental ({note})",
                })

        # If still missing, derive Q2/Q3 from YTD supplemental values
        for end in ends:
            if pd.notna(hist.loc[hist["quarter"] == end, "tax_paid"]).any():
                continue
            if end is None:
                continue
            qi = 4 if end.month == 12 else (end.month // 3)
            if qi not in (2, 3):
                continue
            ytd_map = tax_6m if qi == 2 else tax_9m
            payload = ytd_map.get(end)
            if not payload:
                continue
            ytd_val = payload.get("value")
            if ytd_val is None:
                continue
            prev_end = dt.date(end.year, 3, 31) if qi == 2 else dt.date(end.year, 6, 30)
            prev_val = None
            if prev_end in list(hist["quarter"]):
                prev_val = pd.to_numeric(
                    hist.loc[hist["quarter"] == prev_end, "tax_paid"].iloc[0],
                    errors="coerce",
                )
            if prev_val is None or pd.isna(prev_val):
                prev_payload = tax_3m.get(prev_end)
                if prev_payload:
                    prev_val = prev_payload.get("value")
            if prev_val is None or pd.isna(prev_val):
                continue
            diff = float(ytd_val) - float(prev_val)
            hist.loc[hist["quarter"] == end, "tax_paid"] = diff
            note = payload.get("label") or "cash income tax payments (supplemental)"
            audit_rows.append({
                "metric": "tax_paid",
                "quarter": end,
                "source": "derived_ytd_tax_paid",
                "tag": "CashIncomeTaxPaymentsNet",
                "accn": payload.get("accn"),
                "form": payload.get("form"),
                "filed": parse_date(payload.get("filing_date")),
                "start": None,
                "end": end,
                "unit": "USD",
                "duration_days": None,
                "value": diff,
                "note": f"tax_paid = YTD({qi*3}M) - prior quarter (supplemental; {note})",
            })

    # Tier 2 fallback for COGS / Gross Profit / Op Income from 10-Q/10-K tables
    if sec is not None and submissions is not None and cik_int is not None:
        filing_inventory = _get_filing_shared_inventory()
        with _timed_stage(gaap_stage_timings, "gaap_history.income_statement_fallback", enabled=profile_timings):
            fallback_map, _fb_audit = build_income_statement_fallback(
                sec,
                cik_int,
                submissions,
                max_quarters=max_quarters,
                ticker=ticker,
                filing_inventory=filing_inventory,
                filing_runtime_cache=filing_runtime_cache,
            )
        if fallback_map:
            for end in ends:
                payload = fallback_map.get(end)
                if not payload:
                    continue
                vals = payload.get("values", {})
                labels = payload.get("labels", {})
                for metric in ("cogs", "gross_profit", "op_income"):
                    if metric in hist.columns and pd.isna(hist.loc[hist["quarter"] == end, metric]).any():
                        v = vals.get(metric)
                        if v is None:
                            continue
                        hist.loc[hist["quarter"] == end, metric] = float(v)
                        label_note = labels.get(metric, "")
                        if labels.get("_title_match") == "title_missing":
                            label_note = (label_note + "; title_missing").strip("; ")
                        audit_rows.append({
                            "metric": metric,
                            "quarter": end,
                            "source": "tier2_table",
                            "tag": "10-Q/10-K table",
                            "accn": None,
                            "form": None,
                            "filed": None,
                            "start": None,
                            "end": end,
                            "unit": "USD",
                            "duration_days": None,
                            "value": float(v),
                            "note": f"fallback from income statement table ({label_note})",
                        })
            # Recompute gross_profit/op_income formulas after fallback
            if "gross_profit" in hist.columns and "cogs" in hist.columns:
                for idx, row in hist.iterrows():
                    if pd.isna(row.get("gross_profit")) and pd.notna(row.get("revenue")) and pd.notna(row.get("cogs")):
                        val = float(row["revenue"]) - float(row["cogs"])
                        hist.at[idx, "gross_profit"] = val
                        q = row["quarter"]
                        audit_rows.append({
                            "metric": "gross_profit",
                            "quarter": q,
                            "source": "derived_formula",
                            "value": val,
                            "note": "gross_profit = revenue - cogs (post tier2 fallback)",
                        })
            if "op_income" in hist.columns and "gross_profit" in hist.columns:
                for idx, row in hist.iterrows():
                    if pd.isna(row.get("op_income")) and pd.notna(row.get("gross_profit")):
                        q = row["quarter"]
                        gp = float(row["gross_profit"])
                        pr_opex = _pick_duration_from_facts(opex_facts, q, quarter_index(q) or 0, ["10-Q", "10-K"]) if opex_tag else None
                        if pr_opex is not None and pr_opex.value is not None:
                            val = gp - float(pr_opex.value)
                            hist.at[idx, "op_income"] = val
                            audit_rows.append({
                                "metric": "op_income",
                                "quarter": q,
                                "source": "derived_formula",
                                "value": val,
                                "note": f"op_income = gross_profit - OperatingExpenses ({pr_opex.tag})",
                            })
                            continue
                        pr_sga = _pick_duration_from_facts(sga_facts, q, quarter_index(q) or 0, ["10-Q", "10-K"]) if sga_tag else None
                        pr_rd = _pick_duration_from_facts(
                            rd_facts,
                            q,
                            quarter_index(q) or 0,
                            ["10-Q", "10-K"],
                            metric_name="research_and_development",
                        ) if rd_tag else None
                        if pr_sga is not None and pr_sga.value is not None and pr_rd is not None and pr_rd.value is not None:
                            val = gp - float(pr_sga.value) - float(pr_rd.value)
                            hist.at[idx, "op_income"] = val
                            audit_rows.append({
                                "metric": "op_income",
                                "quarter": q,
                                "source": "derived_formula",
                                "value": val,
                                "note": f"op_income = gross_profit - (SG&A {pr_sga.tag} + R&D {pr_rd.tag})",
                            })

        # Tier 3 fallback from 10-K Quarterly Financial Data (strict)
        missing_qdata_quarters: set[dt.date] = set()
        missing_qdata_by_metric: Dict[str, set[dt.date]] = {}
        for metric in ("revenue", "cogs", "gross_profit", "op_income", "net_income"):
            if metric in hist.columns:
                miss = set([d for d in hist.loc[hist[metric].isna(), "quarter"] if pd.notna(d)])
                if miss:
                    missing_qdata_by_metric[metric] = miss
                    missing_qdata_quarters.update(miss)
        if missing_qdata_quarters:
            target_years = {d.year for d in missing_qdata_quarters}
            target_years |= {y + 1 for y in target_years}
            # also include latest 10-K year for preview visibility
            filing_inventory = _get_filing_shared_inventory()
            latest_10k_year = (
                filing_inventory.get("latest_10k_year")
                if isinstance(filing_inventory, dict)
                else None
            )
            if latest_10k_year:
                target_years.add(latest_10k_year)
                target_years.add(latest_10k_year + 1)
            with _timed_stage(gaap_stage_timings, "gaap_history.quarterly_data_10k_fallback", enabled=profile_timings):
                qdata_map, _qdata_audit, qdata_preview_rows = build_quarterly_data_10k_fallback(
                    sec,
                    cik_int,
                    submissions,
                    max_quarters=max_quarters,
                    target_years=target_years,
                    filing_inventory=filing_inventory,
                    filing_runtime_cache=filing_runtime_cache,
                )
            qfd_preview_rows.extend(qdata_preview_rows)
            if qdata_map:
                for end in ends:
                    payload = qdata_map.get(end)
                    if not payload:
                        continue
                    for metric in ("revenue", "cogs", "gross_profit", "op_income", "net_income"):
                        if metric in hist.columns and pd.isna(hist.loc[hist["quarter"] == end, metric]).any():
                            v = payload.get(metric)
                            if v is None:
                                continue
                            hist.loc[hist["quarter"] == end, metric] = float(v)
                            audit_rows.append({
                                "metric": metric,
                                "quarter": end,
                                "source": "tier3_10k_quarterly_data",
                                "tag": "10-K Quarterly Financial Data",
                                "accn": None,
                                "form": "10-K",
                                "filed": None,
                                "start": None,
                                "end": end,
                                "unit": "USD",
                                "duration_days": None,
                                "value": float(v),
                                "note": "fallback from 10-K Quarterly Financial Data",
                            })
                        elif metric in hist.columns and end in payload:
                            # QFD found but not used (already had value)
                            if metric in missing_qdata_by_metric and end in missing_qdata_by_metric.get(metric, set()):
                                continue
                            v = payload.get(metric)
                            if v is not None:
                                qfd_unused_rows.append({
                                    "quarter": end,
                                    "metric": metric,
                                    "value": float(v),
                                    "reason": "value already present",
                                    "source": "10-K Quarterly Financial Data",
                                })

        # Tier 2b: Q4 derived from 10-K FY minus 10-Q 9M statement tables (strict)
        missing_q4_quarters: set[dt.date] = set()
        for metric in ("cogs", "gross_profit", "op_income"):
            if metric in hist.columns:
                miss = hist.loc[hist[metric].isna(), "quarter"]
                missing_q4_quarters.update([d for d in miss if pd.notna(d) and getattr(d, "month", None) == 12])
        if missing_q4_quarters:
            filing_inventory = _get_filing_shared_inventory()
            with _timed_stage(gaap_stage_timings, "gaap_history.income_statement_ytd_q4_fallback", enabled=profile_timings):
                q4_map, q4_audit = build_income_statement_ytd_q4_fallback(
                    sec,
                    cik_int,
                    submissions,
                    max_quarters=max_quarters,
                    ticker=ticker,
                    target_quarters=missing_q4_quarters,
                    filing_inventory=filing_inventory,
                    filing_runtime_cache=filing_runtime_cache,
                    stage_timings=gaap_stage_timings,
                    profile_timings=profile_timings,
                )
            if q4_map:
                audit_lookup = {(r.get("quarter"), r.get("metric")): r for r in q4_audit}
                for end in ends:
                    if end not in q4_map:
                        continue
                    payload = q4_map.get(end, {})
                    for metric in ("revenue", "cogs", "gross_profit", "op_income", "net_income"):
                        if metric in hist.columns and pd.isna(hist.loc[hist["quarter"] == end, metric]).any():
                            v = payload.get(metric)
                            if v is None:
                                continue
                            hist.loc[hist["quarter"] == end, metric] = float(v)
                            ar = audit_lookup.get((end, metric))
                            if ar:
                                audit_rows.append(ar)
                            else:
                                audit_rows.append({
                                    "metric": metric,
                                    "quarter": end,
                                    "source": "derived_ytd_q4_table",
                                    "value": float(v),
                                    "note": "Q4 = FY - 9M from statement tables",
                                })
                # Recompute gross_profit/op_income after Q4 derivation
                if "gross_profit" in hist.columns and "cogs" in hist.columns:
                    for idx, row in hist.iterrows():
                        if pd.isna(row.get("gross_profit")) and pd.notna(row.get("revenue")) and pd.notna(row.get("cogs")):
                            val = float(row["revenue"]) - float(row["cogs"])
                            hist.at[idx, "gross_profit"] = val
                            q = row["quarter"]
                            audit_rows.append({
                                "metric": "gross_profit",
                                "quarter": q,
                                "source": "derived_formula",
                                "value": val,
                                "note": "gross_profit = revenue - cogs (post Q4 table derivation)",
                            })
                if "op_income" in hist.columns and "gross_profit" in hist.columns:
                    for idx, row in hist.iterrows():
                        if pd.isna(row.get("op_income")) and pd.notna(row.get("gross_profit")):
                            q = row["quarter"]
                            gp = float(row["gross_profit"])
                            pr_opex = _pick_duration_from_facts(opex_facts, q, quarter_index(q) or 0, ["10-Q", "10-K"]) if opex_tag else None
                            if pr_opex is not None and pr_opex.value is not None:
                                val = gp - float(pr_opex.value)
                                hist.at[idx, "op_income"] = val
                                audit_rows.append({
                                    "metric": "op_income",
                                    "quarter": q,
                                    "source": "derived_formula",
                                    "value": val,
                                    "note": f"op_income = gross_profit - OperatingExpenses ({pr_opex.tag})",
                                })
                                continue
                            pr_sga = _pick_duration_from_facts(sga_facts, q, quarter_index(q) or 0, ["10-Q", "10-K"]) if sga_tag else None
                            pr_rd = _pick_duration_from_facts(
                                rd_facts,
                                q,
                                quarter_index(q) or 0,
                                ["10-Q", "10-K"],
                                metric_name="research_and_development",
                            ) if rd_tag else None
                            if pr_sga is not None and pr_sga.value is not None and pr_rd is not None and pr_rd.value is not None:
                                val = gp - float(pr_sga.value) - float(pr_rd.value)
                                hist.at[idx, "op_income"] = val
                                audit_rows.append({
                                    "metric": "op_income",
                                    "quarter": q,
                                    "source": "derived_formula",
                                    "value": val,
                                    "note": f"op_income = gross_profit - (SG&A {pr_sga.tag} + R&D {pr_rd.tag})",
                                })

        # Tier 2 fallback for balance sheet (debt/leases/bank) from 10-Q/10-K tables
        missing_bs_quarters: set[dt.date] = set()
        bs_metrics = ("debt_core", "lease_liabilities", "bank_deposits", "bank_finance_receivables", "total_equity", "goodwill", "intangibles")
        for metric in bs_metrics:
            if metric in hist.columns:
                missing_bs_quarters.update([d for d in hist.loc[hist[metric].isna(), "quarter"] if pd.notna(d)])
        if missing_bs_quarters:
            filing_inventory = _get_filing_shared_inventory()
            with _timed_stage(gaap_stage_timings, "gaap_history.balance_sheet_fallback", enabled=profile_timings):
                bs_map, _bs_audit = build_balance_sheet_fallback_table(
                    sec,
                    cik_int,
                    submissions,
                    max_quarters=max_quarters,
                    target_quarters=missing_bs_quarters,
                    filing_inventory=filing_inventory,
                    filing_runtime_cache=filing_runtime_cache,
                )
            if bs_map:
                for end in ends:
                    payload = bs_map.get(end)
                    if not payload:
                        continue
                    vals = payload.get("values", {})
                    labels = payload.get("labels", {})
                    for metric in bs_metrics:
                        if metric in hist.columns and pd.isna(hist.loc[hist["quarter"] == end, metric]).any():
                            v = vals.get(metric)
                            if v is None:
                                continue
                            hist.loc[hist["quarter"] == end, metric] = float(v)
                            label_note = labels.get(metric, "")
                            audit_rows.append({
                                "metric": metric,
                                "quarter": end,
                                "source": "tier2_table",
                                "tag": "10-Q/10-K balance sheet",
                                "accn": payload.get("accn"),
                                "form": payload.get("form"),
                                "filed": payload.get("filed"),
                                "start": None,
                                "end": end,
                                "unit": "USD",
                                "duration_days": None,
                                "value": float(v),
                                "note": f"balance sheet table fallback ({label_note})",
                            })
                    # If total_debt missing and debt_core available, backfill conservatively
                    if "total_debt" in hist.columns and pd.isna(hist.loc[hist["quarter"] == end, "total_debt"]).any():
                        v = vals.get("debt_core")
                        if v is not None:
                            hist.loc[hist["quarter"] == end, "total_debt"] = float(v)
                            audit_rows.append({
                                "metric": "total_debt",
                                "quarter": end,
                                "source": "tier2_table",
                                "tag": "10-Q/10-K balance sheet",
                                "accn": payload.get("accn"),
                                "form": payload.get("form"),
                                "filed": payload.get("filed"),
                                "start": None,
                                "end": end,
                                "unit": "USD",
                                "duration_days": None,
                                "value": float(v),
                                "note": "total_debt from balance sheet debt_core fallback",
                            })

                # Recompute bank_net_funding where possible
                if "bank_net_funding" in hist.columns:
                    for idx, row in hist.iterrows():
                        if pd.notna(row.get("bank_net_funding")):
                            continue
                        deposits = row.get("bank_deposits")
                        receivables = row.get("bank_finance_receivables")
                        if pd.notna(deposits) and pd.notna(receivables):
                            val = float(deposits) - float(receivables)
                            hist.at[idx, "bank_net_funding"] = val
                            audit_rows.append({
                                "metric": "bank_net_funding",
                                "quarter": row["quarter"],
                                "source": "derived_formula",
                                "value": val,
                                "note": "bank_net_funding = bank_deposits - bank_finance_receivables (post BS fallback)",
                            })

        # Tier 3 OCR fallback from EX-99 images (used only if still missing)
        missing_is_quarters: set[dt.date] = set()
        for metric in ("revenue", "cogs", "gross_profit", "op_income", "net_income"):
            if metric in hist.columns:
                missing_is_quarters.update([d for d in hist.loc[hist[metric].isna(), "quarter"] if pd.notna(d)])
        if missing_is_quarters:
            ex99_inventory = _get_ex99_shared_inventory()
            ocr_map, _ocr_audit = build_income_statement_fallback_ex99_ocr(
                sec,
                cik_int,
                submissions,
                max_quarters=max_quarters,
                ticker=ticker,
                target_quarters=missing_is_quarters,
                ex99_inventory=ex99_inventory,
                ex99_runtime_cache=ex99_runtime_cache,
            )
            if ocr_map:
                for end in ends:
                    payload = ocr_map.get(end)
                    if not payload:
                        continue
                    vals = payload.get("values", {})
                    labels = payload.get("labels", {})
                    for metric in ("revenue", "cogs", "gross_profit", "op_income", "net_income"):
                        if metric in hist.columns and pd.isna(hist.loc[hist["quarter"] == end, metric]).any():
                            v = vals.get(metric)
                            if v is None:
                                continue
                            hist.loc[hist["quarter"] == end, metric] = float(v)
                            label_note = labels.get(metric, "")
                            audit_rows.append({
                                "metric": metric,
                                "quarter": end,
                                "source": "tier3_ex99_ocr",
                                "tag": "EX-99 OCR",
                                "accn": payload.get("accn"),
                                "form": "8-K",
                                "filed": None,
                                "start": None,
                                "end": end,
                                "unit": "USD",
                                "duration_days": None,
                                "value": float(v),
                                "note": f"EX-99 OCR fallback ({label_note})",
                            })

        # Tier 3 OCR fallback for balance sheet (cash/total_debt) if missing
        missing_bs_quarters: set[dt.date] = set()
        for metric in ("cash", "total_debt", "debt_core", "lease_liabilities", "bank_deposits", "bank_finance_receivables"):
            if metric in hist.columns:
                missing_bs_quarters.update([d for d in hist.loc[hist[metric].isna(), "quarter"] if pd.notna(d)])
        if missing_bs_quarters:
            ex99_inventory = _get_ex99_shared_inventory()
            bs_ocr_map, _bs_ocr_audit = build_balance_sheet_fallback_ex99_ocr(
                sec,
                cik_int,
                submissions,
                max_quarters=max_quarters,
                target_quarters=missing_bs_quarters,
                ex99_inventory=ex99_inventory,
                ex99_runtime_cache=ex99_runtime_cache,
            )
            if bs_ocr_map:
                for end in ends:
                    payload = bs_ocr_map.get(end)
                    if not payload:
                        continue
                    vals = payload.get("values", {})
                    labels = payload.get("labels", {})
                    for metric in ("cash", "total_debt", "debt_core", "lease_liabilities", "bank_deposits", "bank_finance_receivables"):
                        if metric in hist.columns and pd.isna(hist.loc[hist["quarter"] == end, metric]).any():
                            v = vals.get(metric)
                            if v is None:
                                continue
                            hist.loc[hist["quarter"] == end, metric] = float(v)
                            label_note = labels.get(metric, "")
                            audit_rows.append({
                                "metric": metric,
                                "quarter": end,
                                "source": "tier3_ex99_ocr",
                                "tag": "EX-99 OCR",
                                "accn": payload.get("accn"),
                                "form": "8-K",
                                "filed": None,
                                "start": None,
                                "end": end,
                                "unit": "USD",
                                "duration_days": None,
                                "value": float(v),
                                "note": f"EX-99 OCR fallback ({label_note})",
                            })
                # Recompute bank_net_funding if OCR filled components
                if "bank_net_funding" in hist.columns:
                    for idx, row in hist.iterrows():
                        if pd.notna(row.get("bank_net_funding")):
                            continue
                        deposits = row.get("bank_deposits")
                        receivables = row.get("bank_finance_receivables")
                        if pd.notna(deposits) and pd.notna(receivables):
                            val = float(deposits) - float(receivables)
                            hist.at[idx, "bank_net_funding"] = val
                            audit_rows.append({
                                "metric": "bank_net_funding",
                                "quarter": row["quarter"],
                                "source": "derived_formula",
                                "value": val,
                                "note": "bank_net_funding = bank_deposits - bank_finance_receivables (post OCR fallback)",
                            })
    audit = pd.DataFrame(audit_rows)
    qfd_preview_df = pd.DataFrame(qfd_preview_rows)
    qfd_unused_df = pd.DataFrame(qfd_unused_rows)
    if not audit.empty:
        audit["source_class"] = audit["source"].apply(_source_class)
        audit["method"] = audit["source"].apply(_source_method)
        audit["qa_severity"] = audit["source"].apply(_source_qa)
        audit["__k"] = list(zip(audit["metric"], audit["quarter"].astype(str)))
        if derived_formula_keys:
            audit = audit[~((audit["source"] == "missing") & (audit["__k"].isin(derived_formula_keys)))]
        # Drop missing rows when a value was filled by fallback/formula
        filled = audit[audit["source"].isin(["direct", "derived_ytd", "derived_ytd_q4", "derived_formula", "tier2_table", "derived_parts"])]
        if not filled.empty:
            filled_keys = set(zip(filled["metric"], filled["quarter"].astype(str)))
            audit = audit[~((audit["source"] == "missing") & (audit["__k"].isin(filled_keys)))]
        audit = audit.drop(columns=["__k"])
    return hist, audit, qfd_preview_df, qfd_unused_df


def build_bridge_q(hist: pd.DataFrame) -> pd.DataFrame:
    if hist is None or hist.empty:
        return pd.DataFrame()
    h = hist.copy()
    cols = [c for c in ["quarter", "revenue", "ebitda", "cfo", "capex", "tax_paid", "interest_paid"] if c in h.columns]
    return h[cols].copy()


def build_qa_checks(df_all: pd.DataFrame, hist: pd.DataFrame, audit: pd.DataFrame | None = None) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    if hist is None or hist.empty or df_all is None or df_all.empty:
        return pd.DataFrame()

    rev_spec = next(s for s in GAAP_SPECS if s.name == "revenue")
    cand = df_all[df_all["tag"].isin(rev_spec.tags)].copy()
    cand = _filter_unit(cand, rev_spec)
    best_tag = choose_best_tag(cand, rev_spec)
    h = hist.copy()
    end_to_fy_fp: Dict[dt.date, Tuple[int, str]] = {}
    fy_to_end: Dict[int, dt.date] = {}
    if best_tag:
        rev = cand[cand["tag"] == best_tag].copy()
        fy_col = "fy_calc" if "fy_calc" in rev.columns else "fy"
        rev = rev[rev[fy_col].notna() & rev["fp"].notna()].copy()
        rev = rev.sort_values([fy_col, "fp", "filed_d"], ascending=[True, True, False])
        for _, r in rev.iterrows():
            if pd.isna(r["end_d"]):
                continue
            end_to_fy_fp.setdefault(r["end_d"], (int(r[fy_col]), str(r["fp"])))
            if str(r["fp"]) == "FY" and int(r[fy_col]) not in fy_to_end:
                fy_to_end[int(r[fy_col])] = r["end_d"]

        def _safe_pair_value(v: Any, idx: int) -> Any:
            if isinstance(v, (tuple, list)) and len(v) > idx:
                return v[idx]
            return None

        mapped = h["quarter"].map(end_to_fy_fp)
        h["fy"] = mapped.map(lambda x: _safe_pair_value(x, 0))
        h["fp"] = mapped.map(lambda x: _safe_pair_value(x, 1))

    # Map quarter -> filed date from audit (for vintage alignment)
    filed_map: Dict[Tuple[str, dt.date], dt.date] = {}
    if audit is not None and not audit.empty and "filed" in audit.columns:
        aud = audit.copy()
        aud["filed_d"] = pd.to_datetime(aud["filed"], errors="coerce").dt.date
        for _, r in aud.iterrows():
            q = r.get("quarter")
            if pd.isna(q):
                continue
            try:
                qd = pd.to_datetime(q).date()
            except Exception:
                continue
            filed = r.get("filed_d")
            if pd.isna(filed):
                continue
            key = (str(r.get("metric")), qd)
            if key not in filed_map or filed > filed_map[key]:
                  filed_map[key] = filed

    # Cash taxes supplemental present but not parsed
    if audit is not None and not audit.empty and "source" in audit.columns:
        miss = audit[audit["source"] == "cash_taxes_missing"].copy()
        if not miss.empty and "tax_paid" in h.columns:
            for _, r in miss.iterrows():
                q = r.get("quarter")
                if pd.isna(q):
                    continue
                try:
                    qd = pd.to_datetime(q).date()
                except Exception:
                    continue
                if pd.notna(h.loc[h["quarter"] == qd, "tax_paid"]).any():
                    continue
                rows.append({
                    "quarter": qd,
                    "metric": "tax_paid",
                    "check": "cash_taxes_missing",
                    "status": "warn",
                    "message": "Cash tax row present in 10-Q supplemental table but not parsed.",
                    "accn": r.get("accn"),
                    "doc": r.get("doc"),
                })

    for spec in GAAP_SPECS:
        if spec.kind != "duration" or spec.name not in h.columns or not spec.tags:
            continue
        if spec.name == "shares_diluted":
            continue
        cand_m = df_all[df_all["tag"].isin(spec.tags)].copy()
        cand_m = _filter_unit(cand_m, spec)
        tag_m = choose_best_tag(cand_m, spec)
        if not tag_m:
            continue
        facts = cand_m[cand_m["tag"] == tag_m].copy()
        for fy, fy_end in fy_to_end.items():
            fy_facts = facts[(facts["end_d"] == fy_end) & facts["start_d"].notna()].copy()
            if fy_facts.empty:
                continue
            fy_facts["dur"] = _duration_days(fy_facts["end_d"], fy_facts["start_d"])
            fy_facts["dur_class"] = fy_facts["dur"].apply(classify_duration)
            fy_facts = fy_facts[fy_facts["dur_class"] == "FY"].copy()
            if fy_facts.empty:
                continue
            # choose FY fact that matches quarter vintage
            q_rows = h[h["fy"] == fy].copy()
            q_rows = q_rows[q_rows["fp"].astype(str).isin({"Q1", "Q2", "Q3", "FY"})]
            filed_dates = []
            for _, qr in q_rows.iterrows():
                fd = filed_map.get((spec.name, pd.to_datetime(qr["quarter"]).date()))
                if fd:
                    filed_dates.append(fd)
            anchor = max(filed_dates) if filed_dates else None
            if anchor is not None:
                grace = dt.timedelta(days=120)
                fy_facts = fy_facts[fy_facts["filed_d"].notna()]
                fy_facts = fy_facts[fy_facts["filed_d"] <= (anchor + grace)]
            if fy_facts.empty:
                rows.append({
                    "quarter": fy_end,
                    "metric": spec.name,
                    "check": "qsum_vs_fy",
                    "status": "skip",
                    "message": "No FY fact within vintage window; tie-out skipped.",
                })
                continue
            fy_fact = fy_facts.sort_values("filed_d", ascending=False).iloc[0]
            if fy_fact is None:
                continue
            fy_rows = h[h["fy"] == fy].copy()
            fps = set(fy_rows["fp"].dropna().astype(str))
            required_fps = {"Q1", "Q2", "Q3", "FY"}
            if not required_fps.issubset(fps):
                continue
            qsum = fy_rows[fy_rows["fp"].astype(str).isin(required_fps)][spec.name].dropna()
            if qsum.shape[0] < 4:
                continue
            sum_val = float(qsum.sum())
            fy_val = float(fy_fact["val"])
            diff = sum_val - fy_val
            diff_pct = diff / fy_val if fy_val != 0 else None
            status = "pass"
            if diff_pct is not None and abs(diff_pct) > 0.10:
                status = "fail"
            elif diff_pct is not None and abs(diff_pct) > 0.02:
                status = "warn"
            rows.append({
                "quarter": fy_end,
                "metric": spec.name,
                "check": "qsum_vs_fy",
                "status": status,
                "message": "Sum of 4 quarters vs FY fact.",
                "sum_quarters": sum_val,
                "fy_fact": fy_val,
                "diff": diff,
                "diff_pct": diff_pct,
            })

    h2 = hist.sort_values("quarter").copy()
    if {"cash", "cfo", "capex", "total_debt"}.issubset(h2.columns):
        h2["cash"] = pd.to_numeric(h2["cash"], errors="coerce")
        h2["cfo"] = pd.to_numeric(h2["cfo"], errors="coerce")
        h2["capex"] = pd.to_numeric(h2["capex"], errors="coerce")
        h2["total_debt"] = pd.to_numeric(h2["total_debt"], errors="coerce")
        h2["delta_cash"] = h2["cash"].diff()
        h2["delta_debt"] = h2["total_debt"].diff()
        h2["residual"] = h2["delta_cash"] - (h2["cfo"] - h2["capex"] + h2["delta_debt"])
        for _, r in h2.iterrows():
            if pd.isna(r["residual"]):
                continue
            base = abs(r["cfo"]) if pd.notna(r["cfo"]) else abs(r["delta_cash"])
            warn_thr = max(0.2 * base, 50_000_000.0)
            fail_thr = max(0.5 * base, 150_000_000.0)
            status = "pass"
            if abs(r["residual"]) > fail_thr:
                status = "fail"
            elif abs(r["residual"]) > warn_thr:
                status = "warn"
            rows.append({
                "quarter": r["quarter"],
                "metric": "cash_identity",
                "check": "cash_identity",
                "status": status,
                "message": "Delta cash vs CFO-capex+Delta debt (approx).",
                "delta_cash": r["delta_cash"],
                "cfo": r["cfo"],
                "capex": r["capex"],
                "delta_debt": r["delta_debt"],
                "residual": r["residual"],
            })

    # Capex sign sanity (capex should be >= 0 in this convention)
    if "capex" in h2.columns:
        for _, r in h2.iterrows():
            cap = pd.to_numeric(r.get("capex"), errors="coerce")
            if pd.notna(cap) and cap < 0:
                rows.append({
                    "quarter": r.get("quarter"),
                    "metric": "capex",
                    "check": "capex_negative",
                    "status": "warn",
                    "message": "Capex is negative; check sign convention.",
                    "capex": cap,
                })

    return pd.DataFrame(rows)


def build_debt_qa_checks(debt_tranches: pd.DataFrame) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    if debt_tranches is None or debt_tranches.empty:
        return pd.DataFrame()

    if "period_match" in debt_tranches.columns:
        for q, grp in debt_tranches.groupby("quarter"):
            pm = grp["period_match"].dropna()
            if pm.empty:
                continue
            if not pm.all():
                rows.append({
                    "quarter": q,
                    "metric": "debt_tranches",
                    "check": "debt_period_match",
                    "status": "warn",
                    "message": "Debt table column date did not match quarter_end; fallback column used.",
                })
    return pd.DataFrame(rows)


def build_interest_qa_checks(hist: pd.DataFrame, audit: Optional[pd.DataFrame]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    if hist is None or hist.empty:
        return pd.DataFrame()
    h = hist.copy()
    h["quarter"] = pd.to_datetime(h["quarter"], errors="coerce")
    h = h[h["quarter"].notna()].sort_values("quarter")
    if "interest_expense_net" in h.columns:
        vals = pd.to_numeric(h["interest_expense_net"], errors="coerce").tolist()
        qs = h["quarter"].tolist()
        for i in range(3, len(vals)):
            window = vals[i - 3 : i + 1]
            if any(pd.isna(v) for v in window):
                continue
            ttm = float(sum(window))
            if ttm <= 0:
                rows.append({
                    "quarter": qs[i],
                    "metric": "interest_expense_net",
                    "check": "interest_expense_zero_or_negative",
                    "status": "warn",
                    "value": ttm,
                    "message": "TTM interest_expense_net is <= 0; P&L interest coverage not meaningful.",
                })

    if audit is not None and not audit.empty:
        a = audit.copy()
        a = a[a["metric"] == "interest_expense_net"].copy()
        if not a.empty:
            a["quarter"] = pd.to_datetime(a["quarter"], errors="coerce")
            a = a[a["quarter"].notna()].sort_values("quarter")
            last_tag = None
            for _, r in a.iterrows():
                tag = r.get("tag")
                if tag and last_tag and tag != last_tag:
                    rows.append({
                        "quarter": r["quarter"],
                        "metric": "interest_expense_net",
                        "check": "interest_tag_switch",
                        "status": "warn",
                        "value": None,
                        "message": f"Interest tag switched: {last_tag} -> {tag}.",
                    })
                if tag:
                    last_tag = tag
    return pd.DataFrame(rows)


def _regression_gate(
    hist: pd.DataFrame,
    audit: pd.DataFrame,
    df_all: pd.DataFrame,
    *,
    ticker: Optional[str] = None,
    cache_dir: Optional[Path] = None,
    debug: bool = False,
    allow_fail: bool = False,
) -> None:
    def _coerce_num(value: Any) -> Optional[float]:
        num = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
        return None if pd.isna(num) else float(num)

    def _coerce_date(value: Any) -> Optional[date]:
        parsed = pd.to_datetime(value, errors="coerce")
        return None if pd.isna(parsed) else parsed.date()

    def _normalization_flags(note: Any) -> Tuple[bool, bool, bool]:
        s = str(note or "").strip().lower()
        scaling = any(tok in s for tok in ("scale", "scaled", "million", "billion", "thousand", "divide", "multipl"))
        sign = any(tok in s for tok in ("absolute value", "used absolute", "sign"))
        period = any(tok in s for tok in ("3m", "quarter", "duration", "annualized", "ytd"))
        return scaling, sign, period

    def _rel_diff(actual: Optional[float], expected: Optional[float]) -> Optional[float]:
        if actual is None or expected is None:
            return None
        denom = max(abs(expected), 1e-9)
        return abs(actual - expected) / denom

    def _is_scaling_issue(actual: Optional[float], expected: Optional[float]) -> bool:
        if actual is None or expected is None:
            return False
        for factor in (1_000.0, 1_000_000.0, 1_000_000_000.0):
            tol = max(1e-6, abs(expected) * 1e-4, abs(actual) * 1e-4)
            if abs((actual * factor) - expected) <= tol or abs(actual - (expected * factor)) <= tol:
                return True
        return False

    def _is_sign_issue(actual: Optional[float], expected: Optional[float]) -> bool:
        if actual is None or expected is None:
            return False
        tol = max(1e-6, max(abs(actual), abs(expected)) * 1e-4)
        return actual * expected < 0 and abs(abs(actual) - abs(expected)) <= tol

    def _is_rounding_issue(actual: Optional[float], expected: Optional[float]) -> bool:
        if actual is None or expected is None:
            return False
        abs_diff = abs(actual - expected)
        rel = _rel_diff(actual, expected)
        return abs_diff <= 1e-3 or (rel is not None and rel <= 1e-4)

    def _nearest_candidate(rows: List[Dict[str, Any]], actual: Optional[float]) -> Optional[Dict[str, Any]]:
        if not rows:
            return None
        if actual is None:
            return rows[0]
        return min(
            rows,
            key=lambda item: abs((_coerce_num(item.get("val")) or 0.0) - actual),
        )

    def _classify_mismatch(
        *,
        actual: Optional[float],
        expected: Optional[float],
        exact_rows: List[Dict[str, Any]],
        accn_end_rows: List[Dict[str, Any]],
        end_rows: List[Dict[str, Any]],
        adjacent_rows: List[Dict[str, Any]],
    ) -> str:
        if _is_scaling_issue(actual, expected):
            return "scaling issue"
        if _is_sign_issue(actual, expected):
            return "sign issue"
        if exact_rows and _is_rounding_issue(actual, expected):
            return "likely benign rounding"
        if not exact_rows and (accn_end_rows or end_rows):
            return "period mismatch"
        if not exact_rows and adjacent_rows:
            return "missing carry-forward"
        return "source parsing mismatch"

    def _fmt_num(value: Any) -> str:
        num = _coerce_num(value)
        if num is None:
            return "n/a"
        return f"{num:,.6f}".rstrip("0").rstrip(".")

    def _write_regression_report(rows: List[Dict[str, Any]]) -> Tuple[Optional[Path], Optional[Path]]:
        if not rows or cache_dir is None:
            return None, None
        out_dir = Path(cache_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        csv_path = out_dir / "regression_gate_failures.csv"
        xlsx_path = out_dir / "regression_gate_failures.xlsx"
        fail_df = pd.DataFrame(rows)
        fail_df.to_csv(csv_path, index=False)
        if debug:
            try:
                fail_df.to_excel(xlsx_path, index=False)
            except Exception:
                xlsx_path = None
        return csv_path, xlsx_path

    failures: List[str] = []
    mismatch_rows: List[Dict[str, Any]] = []
    report_csv_path: Optional[Path] = None
    report_xlsx_path: Optional[Path] = None

    # 1) negative revenue == 0
    if hist is not None and not hist.empty and "revenue" in hist.columns:
        neg_rev = pd.to_numeric(hist["revenue"], errors="coerce").lt(0).sum()
        if neg_rev > 0:
            failures.append(f"negative revenue rows: {int(neg_rev)}")

    # 2) derived_ytd_override == 0
    if audit is not None and not audit.empty and "source" in audit.columns:
        override_ct = audit["source"].astype(str).str.contains("override", na=False).sum()
        if override_ct > 0:
            failures.append(f"derived_ytd_override rows: {int(override_ct)}")

    # 3) direct rows match companyfacts == 100%
    if audit is not None and not audit.empty:
        direct = audit[audit["source"] == "direct"].copy()
        direct = direct[pd.notna(direct["value"])]
        if direct.empty:
            failures.append("direct rows: 0 (expected >0)")
        else:
            df = df_all.copy()
            df["start_d"] = pd.to_datetime(df["start_d"], errors="coerce").dt.date
            df["end_d"] = pd.to_datetime(df["end_d"], errors="coerce").dt.date
            df["start_k"] = df["start_d"].apply(lambda x: x if pd.notna(x) else None)
            df["end_k"] = df["end_d"].apply(lambda x: x if pd.notna(x) else None)
            df["val_key"] = pd.to_numeric(df["val"], errors="coerce").round(6)
            df["accn"] = df["accn"].astype(str)
            df["unit"] = df["unit"].astype(str)
            keyset = set(zip(df["tag"], df["accn"], df["unit"], df["start_k"], df["end_k"], df["val_key"]))
            base_rows = df[["tag", "accn", "unit", "start_k", "end_k", "val", "form", "filed_d", "start_d", "end_d"]].copy()
            base_rows["tag"] = base_rows["tag"].astype(str)
            base_rows["unit"] = base_rows["unit"].astype(str)
            base_rows["accn"] = base_rows["accn"].astype(str)
            full_idx: Dict[Tuple[str, str, str, Optional[date], Optional[date]], List[Dict[str, Any]]] = {}
            accn_end_idx: Dict[Tuple[str, str, str, Optional[date]], List[Dict[str, Any]]] = {}
            end_idx: Dict[Tuple[str, str, Optional[date]], List[Dict[str, Any]]] = {}
            tag_unit_idx: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
            for row in base_rows.to_dict("records"):
                full_key = (row["tag"], row["accn"], row["unit"], row["start_k"], row["end_k"])
                full_idx.setdefault(full_key, []).append(row)
                accn_end_key = (row["tag"], row["accn"], row["unit"], row["end_k"])
                accn_end_idx.setdefault(accn_end_key, []).append(row)
                end_key = (row["tag"], row["unit"], row["end_k"])
                end_idx.setdefault(end_key, []).append(row)
                tag_unit_idx.setdefault((row["tag"], row["unit"]), []).append(row)

            direct["start_d"] = pd.to_datetime(direct["start"], errors="coerce").dt.date
            direct["end_d"] = pd.to_datetime(direct["end"], errors="coerce").dt.date
            direct["start_k"] = direct["start_d"].apply(lambda x: x if pd.notna(x) else None)
            direct["end_k"] = direct["end_d"].apply(lambda x: x if pd.notna(x) else None)
            direct["gate_value"] = pd.to_numeric(direct["value"], errors="coerce")
            if "raw_value" in direct.columns:
                direct["gate_value"] = pd.to_numeric(direct["raw_value"], errors="coerce").combine_first(direct["gate_value"])
            direct["val_key"] = pd.to_numeric(direct["gate_value"], errors="coerce").round(6)
            direct["accn"] = direct["accn"].astype(str)
            direct["unit"] = direct["unit"].astype(str)
            keys = list(zip(direct["tag"], direct["accn"], direct["unit"], direct["start_k"], direct["end_k"], direct["val_key"]))
            matched = sum(1 for k in keys if k in keyset)
            if matched != len(keys):
                unmatched = direct[[k not in keyset for k in keys]].copy()
                for _, row in unmatched.iterrows():
                    tag = str(row.get("tag") or "")
                    accn = str(row.get("accn") or "")
                    unit = str(row.get("unit") or "")
                    start_k = row.get("start_k")
                    end_k = row.get("end_k")
                    actual = _coerce_num(row.get("gate_value"))
                    normalized_value = _coerce_num(row.get("normalized_value"))
                    exact_rows = full_idx.get((tag, accn, unit, start_k, end_k), [])
                    accn_end_rows = accn_end_idx.get((tag, accn, unit, end_k), [])
                    end_rows = end_idx.get((tag, unit, end_k), [])
                    adjacent_rows = []
                    for cand in tag_unit_idx.get((tag, unit), []):
                        cand_end = cand.get("end_k")
                        if cand_end is None or end_k is None or cand_end == end_k:
                            continue
                        if abs((cand_end - end_k).days) <= 100:
                            adjacent_rows.append(cand)
                    expected_row = (
                        _nearest_candidate(exact_rows, actual)
                        or _nearest_candidate(accn_end_rows, actual)
                        or _nearest_candidate(end_rows, actual)
                        or _nearest_candidate(adjacent_rows, actual)
                    )
                    expected = _coerce_num(None if expected_row is None else expected_row.get("val"))
                    abs_diff = None if actual is None or expected is None else abs(actual - expected)
                    rel_diff = _rel_diff(actual, expected)
                    scaling_norm, sign_norm, period_norm = _normalization_flags(row.get("note"))
                    mismatch_rows.append(
                        {
                            "ticker": str(ticker or "").upper(),
                            "metric": row.get("metric"),
                            "quarter": row.get("quarter"),
                            "period_end": end_k,
                            "period_start": start_k,
                            "expected_companyfacts_value": expected,
                            "actual_direct_value": actual,
                            "normalized_value": normalized_value,
                            "abs_diff": abs_diff,
                            "relative_diff": rel_diff,
                            "tag": tag,
                            "accn": accn,
                            "form": row.get("form"),
                            "filed": row.get("filed"),
                            "unit": unit,
                            "source_file": row.get("source_file") or "",
                            "source_sheet": row.get("source_sheet") or "",
                            "source_row": row.get("source_row") or "",
                            "parser_source_type": row.get("source"),
                            "source_choice": row.get("source_choice") or "",
                            "note": row.get("note") or "",
                            "scaling_normalized": scaling_norm,
                            "sign_normalized": sign_norm,
                            "period_normalized": period_norm,
                            "likely_cause": _classify_mismatch(
                                actual=actual,
                                expected=expected,
                                exact_rows=exact_rows,
                                accn_end_rows=accn_end_rows,
                                end_rows=end_rows,
                                adjacent_rows=adjacent_rows,
                            ),
                            "expected_match_level": (
                                "exact"
                                if exact_rows
                                else "same_accn_end"
                                if accn_end_rows
                                else "same_end"
                                if end_rows
                                else "adjacent_period"
                                if adjacent_rows
                                else "none"
                            ),
                            "expected_match_accn": None if expected_row is None else expected_row.get("accn"),
                            "expected_match_form": None if expected_row is None else expected_row.get("form"),
                            "expected_match_filed": None if expected_row is None else expected_row.get("filed_d"),
                            "expected_match_start": None if expected_row is None else expected_row.get("start_d"),
                            "expected_match_end": None if expected_row is None else expected_row.get("end_d"),
                        }
                    )
                report_csv_path, report_xlsx_path = _write_regression_report(mismatch_rows)
                failures.append(f"direct rows not matching companyfacts: {len(keys) - matched} of {len(keys)}")

    # 4) shares_diluted plausibility
    if hist is not None and not hist.empty and "shares_diluted" in hist.columns:
        sh = pd.to_numeric(hist["shares_diluted"], errors="coerce")
        bad = sh.notna() & ((sh <= 0) | (sh >= 1_000_000_000))
        bad_ct = int(bad.sum())
        if bad_ct > 0:
            failures.append(f"shares_diluted out of range: {bad_ct}")

    if failures:
        if mismatch_rows:
            print(
                f"[regression_gate] wrote mismatch report: {report_csv_path}"
                + (f" | {report_xlsx_path}" if report_xlsx_path is not None else ""),
                flush=True,
            )
            preview_n = 10 if debug else 5
            for _, row in pd.DataFrame(mismatch_rows).head(preview_n).iterrows():
                print(
                    "[regression_gate] "
                    f"{row.get('metric')} | {row.get('quarter')} | "
                    f"expected={_fmt_num(row.get('expected_companyfacts_value'))} | "
                    f"actual={_fmt_num(row.get('actual_direct_value'))} | "
                    f"cause={row.get('likely_cause')}",
                    flush=True,
                )
        message = "Regression gate failed: " + "; ".join(failures)
        if report_csv_path is not None:
            message += f" | report={report_csv_path}"
        if allow_fail:
            print(f"[WARN] {message} | continuing because --allow-regression-gate-fail was set.", flush=True)
            return
        raise RuntimeError(message)




def run_pipeline(
    config: PipelineConfig,
    sec_config: SecConfig,
    *,
    ticker: Optional[str] = None,
    cik: Optional[str] = None,
) -> Tuple[pd.DataFrame, ...]:
    # The orchestration layer returns a structured artifact bundle. This wrapper keeps
    # the older tuple-based interface intact so existing callers do not need to know
    # about internal artifact classes.
    from .pipeline_orchestration import run_pipeline_impl

    artifacts = run_pipeline_impl(
        config,
        sec_config,
        ticker=ticker,
        cik=cik,
    )
    return artifacts.as_legacy_tuple()


def write_excel(
    out_path: Path,
    *,
    hist: pd.DataFrame,
    audit: pd.DataFrame,
    needs_review: pd.DataFrame,
    debt_tranches: pd.DataFrame,
    debt_recon: pd.DataFrame,
    adj_metrics: pd.DataFrame,
    adj_breakdown: pd.DataFrame,
    non_gaap_files: pd.DataFrame,
    adj_metrics_relaxed: pd.DataFrame,
    adj_breakdown_relaxed: pd.DataFrame,
    non_gaap_files_relaxed: pd.DataFrame,
    info_log: pd.DataFrame,
    tag_coverage: pd.DataFrame,
    period_checks: pd.DataFrame,
    qa_checks: pd.DataFrame,
    bridge_q: pd.DataFrame,
    manifest_df: pd.DataFrame,
    ocr_log: pd.DataFrame,
    qfd_preview: pd.DataFrame,
    qfd_unused: pd.DataFrame,
    debt_profile: pd.DataFrame,
    debt_tranches_latest: pd.DataFrame,
    debt_maturity: pd.DataFrame,
    debt_credit_notes: pd.DataFrame,
    revolver_df: pd.DataFrame,
    revolver_history: pd.DataFrame,
    debt_buckets: pd.DataFrame,
    slides_segments: pd.DataFrame,
    slides_debt: pd.DataFrame,
    slides_guidance: pd.DataFrame,
    quarter_notes: pd.DataFrame,
    promises: pd.DataFrame,
    promise_progress: pd.DataFrame,
    non_gaap_cred: pd.DataFrame,
    company_overview: Optional[Dict[str, Any]] = None,
    ticker: Optional[str] = None,
    price: Optional[float] = None,
    strictness: str = 'ytd',
    excel_mode: str = 'clean',
    is_rules: Optional[Dict[str, Any]] = None,
    cache_dir: Optional[Path] = None,
    quiet_pdf_warnings: bool = True,
    rebuild_doc_text_cache: bool = False,
    profile_timings: bool = False,
    quarter_notes_audit: bool = False,
    capture_saved_workbook_provenance: bool = True,
    excel_debug_scope: str = 'full',
) -> Any:
    # This wrapper is intentionally mechanical: it documents the dataframe bundle that
    # crosses from pipeline space into workbook-rendering space.
    from .excel_writer import write_excel_from_inputs

    inputs = WorkbookInputs(
        out_path=Path(out_path),
        hist=hist,
        audit=audit,
        needs_review=needs_review,
        debt_tranches=debt_tranches,
        debt_recon=debt_recon,
        adj_metrics=adj_metrics,
        adj_breakdown=adj_breakdown,
        non_gaap_files=non_gaap_files,
        adj_metrics_relaxed=adj_metrics_relaxed,
        adj_breakdown_relaxed=adj_breakdown_relaxed,
        non_gaap_files_relaxed=non_gaap_files_relaxed,
        info_log=info_log,
        tag_coverage=tag_coverage,
        period_checks=period_checks,
        qa_checks=qa_checks,
        bridge_q=bridge_q,
        manifest_df=manifest_df,
        ocr_log=ocr_log,
        qfd_preview=qfd_preview,
        qfd_unused=qfd_unused,
        debt_profile=debt_profile,
        debt_tranches_latest=debt_tranches_latest,
        debt_maturity=debt_maturity,
        debt_credit_notes=debt_credit_notes,
        revolver_df=revolver_df,
        revolver_history=revolver_history,
        debt_buckets=debt_buckets,
        slides_segments=slides_segments,
        slides_debt=slides_debt,
        slides_guidance=slides_guidance,
        quarter_notes=quarter_notes,
        promises=promises,
        promise_progress=promise_progress,
        non_gaap_cred=non_gaap_cred,
        company_overview=company_overview,
        ticker=ticker,
        price=price,
        strictness=strictness,
        excel_mode=excel_mode,
        is_rules=is_rules,
        cache_dir=cache_dir,
        quiet_pdf_warnings=quiet_pdf_warnings,
        rebuild_doc_text_cache=rebuild_doc_text_cache,
        profile_timings=profile_timings,
        quarter_notes_audit=quarter_notes_audit,
        capture_saved_workbook_provenance=capture_saved_workbook_provenance,
        excel_debug_scope=excel_debug_scope,
    )
    return write_excel_from_inputs(inputs)
