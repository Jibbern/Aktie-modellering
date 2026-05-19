"""Core pipeline assembly and stage-cache orchestration.

This module combines SEC facts, local materials, document intelligence, debt parsing,
and summary resolution into a single `PipelineArtifacts` bundle. It is also the main
fine-grained persistence boundary for stage caches such as GAAP history, debt outputs,
local non-GAAP fallback, `doc_intel`, and company overview.
"""
from __future__ import annotations

import datetime as dt
import hashlib
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
try:  # optional in tests, available in the project environment
    from bs4 import BeautifulSoup
except Exception:  # pragma: no cover - dependency guard
    BeautifulSoup = None

from .debt_parser import build_debt_schedule_tier2, build_debt_tranches_tier2, coerce_number
from .doc_intel import build_doc_intel_outputs, extract_pdf_text_cached, validate_quarter_notes
from .legacy_support import (
    PIPELINE_STAGE_CACHE_VERSION,
    _coerce_prev_quarter_end,
    build_bridge_q,
    build_company_overview,
    build_debt_buckets,
    build_debt_credit_notes,
    build_debt_profile,
    build_debt_qa_checks,
    build_gaap_history,
    build_interest_qa_checks,
    build_local_main_revolver_history,
    build_qa_checks,
    build_revolver_availability,
    build_revolver_capacity_map,
    build_revolver_history,
    compute_long_term_debt_instant,
    compute_total_debt_instant,
)
from .metrics import GAAP_SPECS, MetricSpec
from .non_gaap import build_non_gaap_tier3, infer_quarter_end_from_text, parse_adjusted_from_plain_text, strip_html
from .period_resolver import _duration_days, _filter_unit, classify_duration, pick_best_instant, self_check_period_logic
from .pdf_utils import silence_pdfminer_warnings
from .sec_xbrl import SecClient, SecConfig, cik10_from_int, cik_from_ticker, companyfacts_to_df, parse_date
from .validators import info_log_from_audit, needs_review_from_audit, validate_debt_tieout, validate_history
from .cache_layout import preferred_ticker_cache_root_from_base_dir
from .conference_metadata import is_structured_metadata_path, metadata_source_file, parse_metadata_key_values

from .pipeline_qa import (
    build_non_gaap_cred_qa,
    build_promise_qa_checks,
    concat_frames,
    finalize_needs_review,
    finalize_qa_checks,
)
from .pipeline_runtime import (
    PipelineStageCache,
    dataframe_quick_signature,
    material_dirs_signature,
    resolve_pipeline_roots,
    submissions_recent_signature,
    timed_stage,
)
from .pipeline_types import PipelineArtifacts, PipelineConfig
from .source_material_refresh import _looks_preliminary_results_guidance_update


LOCAL_NON_GAAP_FALLBACK_VERSION = 30
LOCAL_NON_GAAP_PDF_PAGE_CACHE_VERSION = 1
DOC_INTEL_BEHAVIOR_VERSION = "v18_anf_source_notes"
COMPANY_OVERVIEW_BEHAVIOR_VERSION = "v9_anf_summary_sanitize"


def _module_code_signature(*relative_names: str) -> str:
    rows: List[str] = []
    base_dir = Path(__file__).resolve().parent
    for rel_name in relative_names:
        try:
            mod_path = (base_dir / rel_name).resolve()
            st = mod_path.stat()
            rows.append(f"{mod_path.name}:{int(st.st_size)}:{int(st.st_mtime)}")
        except Exception:
            rows.append(f"{Path(rel_name).name}:missing")
    if not rows:
        return "none"
    return hashlib.sha1("||".join(rows).encode("utf-8", errors="ignore")).hexdigest()
LOCAL_NON_GAAP_CANONICAL_METRICS: Tuple[str, ...] = ("adj_ebitda", "adj_ebit", "adj_eps", "adj_fcf")


def _local_non_gaap_page_scores(text: str) -> Dict[str, int]:
    t = str(text or "").lower()
    score = {"non_gaap": 0, "segment": 0, "debt": 0, "guidance": 0}
    if "reconciliation of reported net income" in t:
        score["non_gaap"] += 5
    if "reconciliation of reported" in t or "reconciliation of reported consolidated results" in t:
        score["non_gaap"] += 4
    if "reconciliation" in t and "adjusted ebitda" in t:
        score["non_gaap"] += 3
    if "adjusted ebitda" in t and "adjusted ebit" in t:
        score["non_gaap"] += 2
    if "adjusted ebitda" in t and ("net income" in t or "net loss" in t):
        score["non_gaap"] += 2
    if "adjusted diluted earnings per share" in t:
        score["non_gaap"] += 2
    if "schedule of non-gaap financial measures" in t and "adjusted non-gaap" in t:
        score["non_gaap"] += 4
    if "adjusted non-gaap" in t and ("operating income" in t or "net income per diluted share" in t):
        score["non_gaap"] += 3
    if "free cash flow" in t and "capital expenditures" in t:
        score["non_gaap"] += 2
    if "adjusted segment ebit" in t or "reportable segments" in t or "adjusted segment ebitda" in t:
        score["segment"] += 3
    if "net sales by segment" in t or ("americas" in t and "emea" in t and "apac" in t):
        score["segment"] += 3
    if "net sales by brand family" in t or ("abercrombie" in t and "hollister" in t and "comparable sales" in t):
        score["segment"] += 2
    if "comparable sales" in t:
        score["segment"] += 1
    if "sending technology" in t or "presort" in t:
        score["segment"] += 2
    if "debt profile" in t or "credit agreement" in t:
        score["debt"] += 3
    if "revolving credit facility" in t or "aggregate commitments" in t:
        score["debt"] += 2
    debt_markers = (
        "working capital revolver",
        "working capital financing",
        "long-term debt",
        "convertible debt",
        "convertible note",
        "junior mezzanine",
        "term loan",
        "total debt outstanding",
    )
    if any(marker in t for marker in debt_markers):
        score["debt"] += 2
    if "guidance" in t or "outlook" in t:
        score["guidance"] += 2
    if "fy" in t and ("guidance" in t or "outlook" in t):
        score["guidance"] += 1
    if ("adjusted segment" in t or "reportable segments" in t) and "reconciliation of reported" not in t:
        score["non_gaap"] = min(score["non_gaap"], 1)
    return score


def _detect_local_non_gaap_text_scale(text: str) -> float:
    txt = str(text or "")
    if re.search(r"\(\s*\$?\s*0{3}s?\s*\)|\$\s*0{3}s?\b|in\s+\$?0{3}s?", txt, re.I):
        return 1000.0
    if re.search(r"in\s+millions", txt, re.I):
        return 1_000_000.0
    if re.search(r"in\s+thousands", txt, re.I):
        return 1000.0
    return 1.0


def _local_non_gaap_three_month_lines(lines: List[str]) -> List[str]:
    start = None
    end = None
    for i, ln in enumerate(lines):
        if re.search(r"three\s+months\s+ended|quarter\s+ended|thirteen\s+weeks\s+ended", ln, re.I):
            start = i
            continue
        if start is not None and re.search(r"six\s+months|nine\s+months|twelve\s+months|twenty[-\s]?six\s+weeks|thirty[-\s]?nine\s+weeks|fifty[-\s]?two\s+weeks|fifty[-\s]?three\s+weeks|year\s+ended|fiscal\s+year", ln, re.I):
            if i - start <= 3:
                continue
            end = i
            break
    if start is not None:
        return lines[start:end] if end is not None else lines[start:]
    return lines


def _local_non_gaap_years_from_3m_lines(lines: List[str]) -> List[int]:
    years: List[int] = []
    for i, ln in enumerate(lines[:40]):
        if re.search(r"three months|quarter ended|thirteen weeks", ln, re.I):
            yrs = [int(y) for y in re.findall(r"(20\d{2})", ln)]
            if not yrs:
                for j in range(1, 3):
                    if i + j < len(lines):
                        yrs.extend([int(y) for y in re.findall(r"(20\d{2})", lines[i + j])])
            for y in yrs:
                if y not in years:
                    years.append(y)
            if years:
                break
    return years


def _canonical_local_non_gaap_segment_name(segment_in: Any) -> str:
    low = re.sub(r"\s+", " ", str(segment_in or "").strip().lower())
    if not low:
        return ""
    if "sending technology" in low or re.search(r"\bsendtech\b", low):
        return "SendTech Solutions"
    if "presort" in low:
        return "Presort Services"
    if re.search(r"\bamericas?\b", low):
        return "Americas"
    if re.search(r"\bemea\b", low):
        return "EMEA"
    if re.search(r"\bapac\b|asia[- ]pacific|asia pacific", low):
        return "APAC"
    if "total reportable" in low:
        return "Total reportable segments"
    return re.sub(r"\s+", " ", str(segment_in or "").strip())


def _local_non_gaap_amount_values(line_txt: str, *, scale: float) -> List[float]:
    clean = re.sub(r"\(?-?\d+(?:\.\d+)?%\)?", "", str(line_txt or ""))
    values: List[float] = []
    for token in re.findall(r"\(?-?\d{1,3}(?:,\d{3})*(?:\.\d+)?\)?", clean):
        value = coerce_number(token)
        if value is None:
            continue
        if 1900 <= float(value) <= 2100 and len(str(int(value))) == 4:
            continue
        values.append(float(value) * scale)
    return values


def _pick_local_non_gaap_values_by_year(
    values: List[float],
    years: List[int],
    q_end: Optional[dt.date],
    count: int,
) -> Optional[List[float]]:
    if len(values) < count:
        return None
    if years and q_end is not None and len(values) >= count * 2:
        year = int(q_end.year)
        if year == years[0]:
            return values[:count]
        if len(years) > 1 and year == years[1]:
            return values[count : count * 2]
    return values[:count]


def _parse_local_non_gaap_segment_rows_from_text(
    text: str,
    q_end: Optional[dt.date],
) -> List[Dict[str, Any]]:
    if not text or q_end is None:
        return []
    lines = [re.sub(r"\s+", " ", ln).strip() for ln in str(text or "").splitlines() if ln.strip()]
    lines_3m = _local_non_gaap_three_month_lines(lines)
    years = _local_non_gaap_years_from_3m_lines(lines_3m)
    scale = _detect_local_non_gaap_text_scale(text)
    page_low = str(text or "").lower()
    is_revenue_schedule = "business segment revenue" in page_low or (
        "total revenue" in page_low and "adjusted segment" not in page_low
    )
    is_anf_segment_schedule = "net sales by segment" in page_low or (
        "americas" in page_low and "emea" in page_low and "apac" in page_low and "net sales" in page_low
    )
    period_type = "quarter"
    if re.search(r"\b(fifty[- ]two|fifty[- ]three)\s+weeks?\s+ended\b|\bfiscal years?\s+ended\b|\byear\s+ended\b", page_low):
        period_type = "annual"
    elif re.search(r"\b(twenty[- ]six|thirty[- ]nine)\s+weeks?\s+ended\b", page_low):
        period_type = "ytd"
    rows: List[Dict[str, Any]] = []

    for line in lines_3m:
        low = line.lower()
        segment = ""
        if "sending technology" in low or re.search(r"\bsendtech\b", low):
            segment = "SendTech Solutions"
        elif "presort" in low:
            segment = "Presort Services"
        elif "total reportable" in low:
            segment = "Total reportable segments"
        elif re.match(r"^americas\b", low):
            segment = "Americas"
        elif re.match(r"^emea\b", low):
            segment = "EMEA"
        elif re.match(r"^apac\b", low):
            segment = "APAC"
        if not segment:
            continue

        values = _local_non_gaap_amount_values(line, scale=scale)
        if is_anf_segment_schedule and segment in {"Americas", "EMEA", "APAC"}:
            picked = _pick_local_non_gaap_values_by_year(values, years, q_end, 1)
            if picked:
                rows.append(
                    {
                        "quarter": q_end,
                        "segment": segment,
                        "metric": "revenue",
                        "value": picked[0],
                        "unit": "USD",
                        "period_type": period_type,
                        "source_period_label": period_type,
                    }
                )
            pct_vals: List[float] = []
            for tok in re.findall(r"\(?-?\d+(?:\.\d+)?\)?\s*%", line):
                val = coerce_number(tok.replace("%", ""))
                if val is not None:
                    pct_vals.append(float(val))
            if pct_vals:
                comp_val = pct_vals[-1] if "comparable sales" in page_low and len(pct_vals) >= 2 else pct_vals[0]
                rows.append(
                    {
                        "quarter": q_end,
                        "segment": segment,
                        "metric": "comparable_sales",
                        "value": comp_val / 100.0,
                        "unit": "%",
                        "period_type": period_type,
                        "source_period_label": period_type,
                    }
                )
            continue
        if is_revenue_schedule and segment != "Total reportable segments":
            picked = _pick_local_non_gaap_values_by_year(values, years, q_end, 1)
            if picked:
                rows.append(
                    {
                        "quarter": q_end,
                        "segment": segment,
                        "metric": "revenue",
                        "value": picked[0],
                        "unit": "USD",
                        "period_type": period_type,
                        "source_period_label": period_type,
                    }
                )
            continue

        picked3 = _pick_local_non_gaap_values_by_year(values, years, q_end, 3)
        if not picked3:
            continue
        rows.append({"quarter": q_end, "segment": segment, "metric": "adj_segment_ebit", "value": picked3[0], "unit": "USD", "period_type": period_type, "source_period_label": period_type})
        rows.append({"quarter": q_end, "segment": segment, "metric": "adj_segment_da", "value": picked3[1], "unit": "USD", "period_type": period_type, "source_period_label": period_type})
        rows.append({"quarter": q_end, "segment": segment, "metric": "adj_segment_ebitda", "value": picked3[2], "unit": "USD", "period_type": period_type, "source_period_label": period_type})
    return rows


def _local_non_gaap_segment_row_score(row: Dict[str, Any]) -> Tuple[float, float]:
    doc = str(row.get("doc") or "").lower()
    source = str(row.get("source") or "").lower()
    page = row.get("page")
    value = pd.to_numeric(row.get("value"), errors="coerce")
    value_abs = abs(float(value)) if pd.notna(value) else 0.0
    score = 0.0
    if "earnings_release" in source:
        score += 30.0
    if "earnings_presentation" in source or "earnings_presentation" in doc:
        score += 40.0
    if "annual_reports" in source or "annual_report" in doc:
        score -= 80.0
    if doc.endswith(".pdf"):
        score += 45.0
    if pd.notna(page):
        score += 15.0
    if "financial_statement" in source or "financial_statement" in doc:
        score -= 20.0
    if value_abs >= 1_000_000.0:
        score += 8.0
    elif value_abs and value_abs < 750_000.0:
        score -= 12.0
    if "q1_2026_earnings_release" in doc:
        score += 20.0
    return score, value_abs


def _dedupe_local_non_gaap_segment_rows(rows: pd.DataFrame) -> pd.DataFrame:
    if rows is None or rows.empty:
        return pd.DataFrame()
    df = rows.copy()
    if "quarter" not in df.columns or "segment" not in df.columns or "metric" not in df.columns:
        return df
    df["quarter"] = pd.to_datetime(df["quarter"], errors="coerce")
    df["segment"] = df["segment"].map(_canonical_local_non_gaap_segment_name)
    df["metric"] = df["metric"].astype(str).str.strip()
    df["value"] = pd.to_numeric(df.get("value"), errors="coerce")
    df = df[df["quarter"].notna() & df["segment"].astype(bool) & df["metric"].astype(bool) & df["value"].notna()].copy()
    df = df[~(df["metric"].str.lower().eq("revenue") & (df["value"] <= 0))].copy()
    retail_segments = {"Americas", "EMEA", "APAC"}
    df = df[
        ~(
            df["metric"].str.lower().eq("revenue")
            & df["segment"].isin(retail_segments)
            & (df["value"].abs() < 750_000.0)
        )
    ].copy()
    df = df[
        ~(
            df["metric"].str.lower().isin({"adj_segment_ebit", "adj_segment_da", "adj_segment_ebitda"})
            & (df["value"].abs() < 750_000.0)
        )
    ].copy()
    if df.empty:
        return df
    if "period_type" not in df.columns:
        df["period_type"] = "quarter"
    df["period_type"] = df["period_type"].astype(str).str.strip().str.lower().replace({"": "quarter"})
    scores = [_local_non_gaap_segment_row_score(row) for row in df.to_dict("records")]
    df["_source_score"] = [score for score, _abs_value in scores]
    df["_abs_value"] = [_abs_value for _score, _abs_value in scores]
    df = (
        df.sort_values(["quarter", "period_type", "segment", "metric", "_source_score", "_abs_value"], ascending=[True, True, True, True, False, False])
        .drop_duplicates(subset=["quarter", "period_type", "segment", "metric"], keep="first")
        .drop(columns=["_source_score", "_abs_value"], errors="ignore")
        .reset_index(drop=True)
    )
    return df


def _dedupe_slides_guidance_rows(rows: pd.DataFrame) -> pd.DataFrame:
    if rows is None or rows.empty:
        return pd.DataFrame() if rows is None else rows
    df = rows.copy()
    if "quarter" in df.columns:
        df["quarter"] = pd.to_datetime(df["quarter"], errors="coerce")
    for col in ("period_label", "metric_hint", "numbers", "doc"):
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].astype(str).str.strip()
    key_cols = [c for c in ("quarter", "period_label", "metric_hint", "numbers", "doc") if c in df.columns]
    if not key_cols:
        return df.reset_index(drop=True)
    return df.drop_duplicates(subset=key_cols, keep="first").reset_index(drop=True)


def _local_non_gaap_has_financial_statement_files(base_dir: Path, ticker: str = "") -> bool:
    candidate_dirs = [base_dir / "financial_statement"]
    ticker_u = str(ticker or "").strip().upper()
    if ticker_u:
        candidate_dirs.extend(
            [
                base_dir / f"{ticker_u}-10K",
                base_dir / f"{ticker_u}_10K",
                base_dir / f"{ticker_u} 10K",
            ]
        )
    for folder in candidate_dirs:
        if not folder.exists() or not folder.is_dir():
            continue
        try:
            if any(p.is_file() and p.suffix.lower() in {".txt", ".htm", ".html", ".pdf"} for p in folder.iterdir()):
                return True
        except Exception:
            continue
    return False


def _local_non_gaap_debt_source_allowed(src_name: str, *, has_financial_statement_files: bool) -> bool:
    # Annual-report debt tables are only a fallback when a better statement-specific
    # local source is absent. Once `financial_statement` files exist, those should win.
    if str(src_name or "").strip().lower() == "annual_reports":
        return not has_financial_statement_files
    return True


def _local_non_gaap_actuals_allowed_for_source(src_name: str, path_name: str, text: str) -> bool:
    if str(src_name or "").strip().lower() != "press_release":
        return True
    return not _looks_preliminary_results_guidance_update(f"{path_name or ''} {text or ''}")


def _limit_recent_financial_statement_debt_rows(
    df: pd.DataFrame,
    *,
    max_recent_quarters: int = 6,
) -> pd.DataFrame:
    if df is None or df.empty or "quarter" not in df.columns or max_recent_quarters <= 0:
        return df
    out = df.copy()
    if "source" in out.columns:
        source_mask = out["source"].astype(str).str.lower().eq("financial_statement")
    else:
        source_mask = pd.Series([True] * len(out), index=out.index)
    if not bool(source_mask.any()):
        return out
    q_series = pd.to_datetime(out.loc[source_mask, "quarter"], errors="coerce")
    valid_q = sorted({pd.Timestamp(v) for v in q_series.dropna()})
    if len(valid_q) <= max_recent_quarters:
        return out
    keep_q = set(valid_q[-max_recent_quarters:])
    keep_mask = ~source_mask
    keep_mask.loc[source_mask] = pd.to_datetime(out.loc[source_mask, "quarter"], errors="coerce").isin(keep_q)
    trimmed = out.loc[keep_mask].copy()
    return trimmed.reset_index(drop=True)


def _drop_financial_statement_debt_rows_covered_by_slides(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "source" not in df.columns or "quarter" not in df.columns:
        return df
    out = df.copy()
    source = out["source"].astype(str).str.lower()
    slide_q = {
        pd.Timestamp(v)
        for v in pd.to_datetime(out.loc[source.eq("slides"), "quarter"], errors="coerce").dropna()
    }
    if not slide_q:
        return out
    fs_mask = source.eq("financial_statement")
    fs_q = pd.to_datetime(out.loc[fs_mask, "quarter"], errors="coerce")
    keep_mask = ~fs_mask
    keep_mask.loc[fs_mask] = ~fs_q.isin(slide_q)
    trimmed = out.loc[keep_mask].copy()
    return trimmed.reset_index(drop=True)


def _parse_financial_statement_debt_table_html(path_in: Path, q_end: Optional[dt.date]) -> List[Dict[str, Any]]:
    try:
        tables = pd.read_html(str(path_in))
    except Exception:
        return []

    def _norm_text(value: Any) -> str:
        txt = str(value or "").strip()
        return "" if txt.lower() == "nan" else txt

    best_rows: List[Dict[str, Any]] = []
    for df in tables:
        if df is None or df.empty:
            continue
        candidate_rows: List[Dict[str, Any]] = []
        pending_debt_group = ""
        for _, row in df.fillna("").iterrows():
            values = [_norm_text(v) for v in row.tolist()]
            nonempty = [v for v in values if v]
            if not nonempty:
                continue
            lead_cells = [v for v in values[:3] if v]
            label = lead_cells[0] if lead_cells else nonempty[0]
            low = label.lower()
            if low.startswith("corporate"):
                pending_debt_group = ""
                continue
            if low.startswith("green plains"):
                has_amount = any(
                    coerce_number(cell) is not None
                    and abs(float(coerce_number(cell) or 0.0)) > 0
                    for cell in nonempty[1:]
                    if "%" not in str(cell).lower()
                )
                if not has_amount:
                    pending_debt_group = label
                    continue
            if low.startswith(
                (
                    "total book value",
                    "unamortized",
                    "less:",
                    "total long-term debt",
                    "lease liabilities",
                    "year ending",
                    "thereafter",
                    "total",
                )
            ):
                continue
            is_other = low == "other"
            is_debt_row = is_other or bool(
                re.search(
                    r"(convertible\s+notes?\s+due|notes?\s+due|term\s+loan\s+due|tallgrass\s+term\s+loan\s+due|mezzanine\s+notes?\s+due)",
                    low,
                )
            )
            if not is_debt_row:
                continue
            display_label = label
            if (
                pending_debt_group
                and "tallgrass term loan" in low
                and not low.startswith(pending_debt_group.lower())
            ):
                display_label = f"{pending_debt_group} {label}"
                pending_debt_group = ""
            nums: List[float] = []
            for cell in nonempty[1:]:
                cell_low = cell.lower()
                if "%" in cell_low or "sofr" in cell_low or "libor" in cell_low or "interest rate" in cell_low:
                    continue
                if cell in {"—", "–", "-"}:
                    continue
                num = coerce_number(cell)
                if num is None:
                    continue
                if abs(float(num)) <= 0:
                    continue
                nums.append(float(num))
            if not nums:
                continue
            maturity_match = re.search(r"\b(20\d{2})\b", low)
            candidate_rows.append(
                {
                    "quarter": q_end,
                    "tranche": display_label[:180],
                    "amount": nums[0] * 1000.0,
                    "maturity_year": int(maturity_match.group(1)) if maturity_match else None,
                    "unit": "USD",
                    "is_table_total": False,
                    "asof_col_idx": 0,
                    "asof_match_found": bool(q_end is not None),
                }
            )
        if len(candidate_rows) > len(best_rows):
            best_rows = candidate_rows
    return best_rows


def _parse_local_non_gaap_header_dates(text: str) -> List[dt.date]:
    txt = str(text or "")
    out: List[dt.date] = []
    seen: set[dt.date] = set()

    def _append_date(value: Optional[dt.date]) -> None:
        if value is None or value in seen:
            return
        seen.add(value)
        out.append(value)

    for token in re.findall(r"(?:0?[1-9]|1[0-2])[/-](?:0?[1-9]|[12]\d|3[01])[/-](?:\d{2}|\d{4})", txt):
        m = re.match(r"^\s*(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\s*$", token)
        if not m:
            continue
        mm = int(m.group(1))
        dd = int(m.group(2))
        yy_raw = int(m.group(3))
        yy = yy_raw if yy_raw >= 100 else (2000 + yy_raw if yy_raw <= 69 else 1900 + yy_raw)
        try:
            _append_date(dt.date(yy, mm, dd))
        except Exception:
            continue

    month_pat = re.compile(
        r"(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|"
        r"Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\.?\s+(\d{1,2}),?\s*(\d{4})",
        re.I,
    )
    for match in month_pat.finditer(txt):
        month_txt = str(match.group(1) or "").replace(".", "").strip().lower()
        if month_txt == "sept":
            month_txt = "sep"
        day_num = int(match.group(2))
        year_num = int(match.group(3))
        try:
            parsed = pd.to_datetime(f"{month_txt} {day_num} {year_num}", errors="coerce")
        except Exception:
            parsed = pd.NaT
        if pd.notna(parsed):
            _append_date(pd.Timestamp(parsed).date())

    return out


def _infer_local_non_gaap_period_end_from_name(name: str) -> Optional[dt.date]:
    name_txt = str(name or "")
    m = re.search(r"(20\d{2})[-_]?([01]\d)[-_]?([0-3]\d)", name_txt)
    if m:
        try:
            d = dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            return _coerce_prev_quarter_end(d)
        except Exception:
            pass
    m2 = re.search(r"Q([1-4])\s*(20\d{2})", name_txt, re.IGNORECASE)
    if m2:
        q = int(m2.group(1))
        y = int(m2.group(2))
        return dt.date(y, 3 * q, 30 if q in (2, 3) else 31)
    m3 = re.search(r"(20\d{2})\s*[_-]?Q([1-4])", name_txt, re.IGNORECASE)
    if m3:
        y = int(m3.group(1))
        q = int(m3.group(2))
        return dt.date(y, 3 * q, 30 if q in (2, 3) else 31)
    low = name_txt.lower()
    if re.search(r"(annual[_\s-]?report|10k|10-k|fy20\d{2})", low):
        year_match = re.search(r"(20\d{2})", name_txt)
        if year_match:
            try:
                return dt.date(int(year_match.group(1)), 12, 31)
            except Exception:
                return None
    return None


def _retail_fiscal_aliases_from_history(hist_in: pd.DataFrame) -> Dict[dt.date, dt.date]:
    aliases: Dict[dt.date, dt.date] = {}
    if hist_in is None or hist_in.empty or "quarter" not in hist_in.columns:
        return aliases
    q_series = pd.to_datetime(hist_in["quarter"], errors="coerce")
    for qv in q_series.dropna():
        qd = pd.Timestamp(qv).date()
        if qd.month in (1, 2):
            aliases[dt.date(qd.year - 1, 12, 31)] = qd
        elif qd.month in (4, 5):
            aliases[dt.date(qd.year, 3, 31)] = qd
        elif qd.month in (7, 8):
            aliases[dt.date(qd.year, 6, 30)] = qd
        elif qd.month in (10, 11):
            aliases[dt.date(qd.year, 9, 30)] = qd
    return aliases


def _anf_fiscal_period_from_date(qd: dt.date) -> Optional[Tuple[int, int]]:
    if qd.month in (1, 2):
        return int(qd.year) - 1, 4
    if qd.month in (4, 5):
        return int(qd.year), 1
    if qd.month in (7, 8):
        return int(qd.year), 2
    if qd.month in (10, 11):
        return int(qd.year), 3
    return None


def _anf_fiscal_periods_from_history(hist_in: pd.DataFrame) -> Dict[Tuple[int, int], dt.date]:
    out: Dict[Tuple[int, int], dt.date] = {}
    if hist_in is None or hist_in.empty or "quarter" not in hist_in.columns:
        return out
    for qv in pd.to_datetime(hist_in["quarter"], errors="coerce").dropna():
        qd = pd.Timestamp(qv).date()
        fq = _anf_fiscal_period_from_date(qd)
        if fq is not None:
            out[fq] = qd
    return out


def _anf_line_amount_values(line_txt: str, *, scale: float) -> List[float]:
    clean = re.sub(r"\(?-?\d+(?:\.\d+)?\s*%\)?", "", str(line_txt or ""))
    values: List[float] = []
    for token in re.findall(r"\(?\s*\$?\s*-?\d[\d,]*(?:\.\d+)?\s*\)?", clean):
        value = coerce_number(token)
        if value is None:
            continue
        try:
            v = float(value)
        except Exception:
            continue
        if 1900 <= abs(v) <= 2100 and len(str(int(abs(v)))) == 4:
            continue
        values.append(v * scale)
    return values


def _anf_dedup_cells(values: List[Any]) -> str:
    out: List[str] = []
    for value in values:
        if value is None or pd.isna(value):
            continue
        text = re.sub(r"\s+", " ", str(value)).strip()
        if not text or text.lower() == "nan":
            continue
        if out and out[-1] == text:
            continue
        out.append(text)
    return " ".join(out)


def _anf_html_table_lines(path_in: Path) -> List[str]:
    lines: List[str] = []
    try:
        tables = pd.read_html(str(path_in))
    except Exception:
        tables = []
    for table in tables:
        try:
            for _, row in table.iterrows():
                line = _anf_dedup_cells(row.tolist())
                if line:
                    lines.append(line)
        except Exception:
            continue
    if lines:
        return lines
    if BeautifulSoup is None:
        return []
    try:
        soup = BeautifulSoup(path_in.read_text(encoding="utf-8", errors="ignore"), "html.parser")
    except Exception:
        return []
    for tr in soup.find_all("tr"):
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        line = _anf_dedup_cells(cells)
        if line:
            lines.append(line)
    return lines


def _anf_extract_material_lines(
    path_in: Path,
    *,
    cache_root: Path,
    rebuild_cache: bool,
    quiet_pdf_warnings: bool,
) -> Tuple[str, List[str]]:
    suffix = path_in.suffix.lower()
    if suffix == ".pdf":
        text = extract_pdf_text_cached(
            path_in,
            cache_root=cache_root,
            rebuild_cache=rebuild_cache,
            quiet_pdf_warnings=quiet_pdf_warnings,
        )
        lines = [re.sub(r"\s+", " ", ln).strip() for ln in str(text or "").splitlines() if ln.strip()]
        return str(text or ""), lines
    if suffix in {".htm", ".html"}:
        lines = _anf_html_table_lines(path_in)
        try:
            raw = path_in.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            raw = ""
        if not lines and raw:
            text = strip_html(raw)
            lines = [re.sub(r"\s+", " ", ln).strip() for ln in str(text or "").splitlines() if ln.strip()]
        return "\n".join(lines) if lines else strip_html(raw), lines
    if suffix == ".txt":
        text = path_in.read_text(encoding="utf-8", errors="ignore")
        lines = [re.sub(r"\s+", " ", ln).strip() for ln in text.splitlines() if ln.strip()]
        return text, lines
    return "", []


def _anf_statement_three_month_lines(lines: List[str]) -> List[str]:
    if not lines:
        return []
    start = None
    saw_statement_heading = False
    for i, line in enumerate(lines):
        low = re.sub(r"\s+", " ", str(line or "")).strip().lower()
        if "condensed consolidated statements of operations" in low:
            saw_statement_heading = True
            continue
        if saw_statement_heading and re.search(r"\bthirteen\s+weeks?\s+ended\b", low):
            start = i
            break
    if start is None:
        return _local_non_gaap_three_month_lines(lines)
    end = len(lines)
    for j in range(start + 1, len(lines)):
        low = re.sub(r"\s+", " ", str(lines[j] or "")).strip().lower()
        if j > start + 8 and (
            "reporting and use of gaap" in low
            or "schedule of non-gaap financial measures" in low
            or "reconciliation of" in low
            or "condensed consolidated statements of operations" in low
            or re.search(r"\bfifty[- ](?:two|three)\s+weeks?\s+ended\b", low)
        ):
            end = j
            break
    return lines[start:end]


def _parse_anf_statement_values_from_lines(lines: List[str], *, scale: float) -> Dict[str, float]:
    if not lines:
        return {}
    block = _anf_statement_three_month_lines(lines)
    out: Dict[str, float] = {}
    eps_header_seen = False
    shares_header_seen = False
    for line in block:
        low = re.sub(r"\s+", " ", line).strip().lower()
        if not low:
            continue
        if "net income per share attributable" in low or "net income per diluted share attributable" in low:
            eps_header_seen = True
            if "per diluted share" in low:
                raw_nums = [
                    float(x.replace(",", ""))
                    for x in re.findall(r"\$?\s*([0-9]{1,3}(?:\.[0-9]+)?)", line)
                    if not re.fullmatch(r"20\d{2}", x)
                ]
                if raw_nums:
                    out["eps_diluted"] = float(raw_nums[0])
            continue
        if "weighted-average shares outstanding" in low:
            shares_header_seen = True
            continue
        if eps_header_seen and re.match(r"^diluted\b", low):
            raw_nums = [
                float(x.replace(",", ""))
                for x in re.findall(r"\$?\s*([0-9]{1,3}(?:\.[0-9]+)?)", line)
                if not re.fullmatch(r"20\d{2}", x)
            ]
            if raw_nums:
                out["eps_diluted"] = float(raw_nums[0])
            eps_header_seen = False
            continue
        if shares_header_seen and re.match(r"^diluted\b", low):
            nums_sh = _anf_line_amount_values(line, scale=scale)
            if nums_sh:
                out["shares_diluted"] = float(nums_sh[0])
            shares_header_seen = False
            continue
        nums = _anf_line_amount_values(line, scale=scale)
        if not nums:
            continue
        value = nums[0]
        if re.match(r"^net sales\b", low) and "constant currency" not in low:
            out["revenue"] = value
        elif re.match(r"^cost of sales\b", low):
            out["cogs"] = value
        elif re.match(r"^gross profit\b", low):
            out["gross_profit"] = value
        elif re.match(r"^operating income(?:\s|$)", low) and "adjusted" not in low:
            out["op_income"] = value
        elif "net income attributable" in low and "per share" not in low:
            out["net_income"] = value
        elif re.match(r"^net income(?:\s|$)", low) and "per share" not in low and "attributable" not in low:
            out.setdefault("net_income", value)
    if "gross_profit" not in out and out.get("revenue") is not None and out.get("cogs") is not None:
        out["gross_profit"] = float(out["revenue"]) - float(out["cogs"])
    ebitda = _parse_anf_reconciliation_block_values(lines, duration_re=r"\bthirteen\s+weeks?\s+ended\b", scale=scale)
    if ebitda.get("ebitda") is not None:
        out["ebitda"] = float(ebitda["ebitda"])
    return out


def _parse_anf_cash_flow_ytd_from_lines(lines: List[str], *, scale: float) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for line in lines:
        low = re.sub(r"\s+", " ", line).strip().lower()
        if not low:
            continue
        nums = _anf_line_amount_values(line, scale=scale)
        if not nums:
            continue
        if "net cash" in low and "operating activities" in low and ("provided" in low or "used" in low):
            out["cfo_ytd"] = nums[0]
        elif "purchases of property and equipment" in low:
            out["capex_ytd"] = abs(nums[0])
    return out


def _parse_anf_balance_sheet_values_from_lines(lines: List[str], *, scale: float) -> Dict[str, float]:
    out: Dict[str, float] = {}
    saw_lease_liabilities = False
    saw_conventional_debt = False
    for line in lines:
        low = re.sub(r"\s+", " ", str(line or "")).strip().lower()
        if not low:
            continue
        nums = _anf_line_amount_values(line, scale=scale)
        if not nums:
            continue
        first = float(nums[0])
        if re.match(r"^cash and equivalents\b", low) and "restricted cash" not in low:
            out["cash"] = first
        elif re.match(r"^marketable securities\b", low):
            out["marketable_securities"] = first
        elif re.match(r"^inventories\b", low):
            out["inventory"] = first
        elif "short-term portion of operating lease liabilities" in low:
            out["lease_liabilities_current"] = first
            saw_lease_liabilities = True
        elif "long-term portion of operating lease liabilities" in low:
            out["lease_liabilities_noncurrent"] = first
            saw_lease_liabilities = True
        elif re.search(r"\b(senior secured notes|senior notes|term loan|long-term debt|borrowings|revolving credit)\b", low):
            if first > 0:
                saw_conventional_debt = True
    if saw_lease_liabilities and not saw_conventional_debt and "cash" in out:
        out["debt_core"] = 0.0
    return out


def _parse_anf_reconciliation_block_values(lines: List[str], *, duration_re: str, scale: float) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for idx, line in enumerate(lines):
        low = re.sub(r"\s+", " ", str(line or "")).strip().lower()
        if not re.search(duration_re, low, re.I):
            continue
        prev_blob = " ".join(lines[max(0, idx - 5) : idx + 1]).lower()
        if "reconciliation of ebitda" not in prev_blob:
            continue
        for sub_line in lines[idx + 1 : min(len(lines), idx + 28)]:
            sub_low = re.sub(r"\s+", " ", str(sub_line or "")).strip().lower()
            if idx + 1 < len(lines) and re.search(r"\b(thirteen|twenty-six|thirty-nine|fifty-two|fifty-three)\s+weeks?\s+ended\b", sub_low):
                break
            nums = _anf_line_amount_values(sub_line, scale=scale)
            if not nums:
                continue
            if re.match(r"^adjusted ebitda\b", sub_low):
                out["adj_ebitda"] = float(nums[0])
            elif re.match(r"^ebitda\b", sub_low):
                out["ebitda"] = float(nums[0])
            elif "litigation settlement" in sub_low:
                out["litigation_settlement_adjustment"] = float(nums[0])
        if out:
            return out
    return out


def _parse_anf_non_gaap_schedule_values(lines: List[str], *, duration_re: str, scale: float) -> Dict[str, float]:
    out: Dict[str, float] = {}
    active = False
    for idx, line in enumerate(lines):
        low = re.sub(r"\s+", " ", str(line or "")).strip().lower()
        if "schedule of non-gaap financial measures" in low:
            window = " ".join(lines[idx : min(len(lines), idx + 8)]).lower()
            active = bool(re.search(duration_re, window, re.I))
            continue
        if not active:
            continue
        if idx > 0 and (
            "reconciliation of constant currency" in low
            or "reconciliation of ebitda" in low
            or (low.startswith("abercrombie & fitch co.") and out)
        ):
            break
        nums = _anf_line_amount_values(line, scale=scale)
        if re.match(r"^operating income\b", low) and len(nums) >= 2:
            out["adj_ebit"] = float(nums[-1])
        elif "litigation settlement" in low and nums:
            out["litigation_settlement_adjustment"] = float(nums[0])
        elif "net income per diluted share attributable" in low:
            raw_nums = [
                float(x.replace(",", ""))
                for x in re.findall(r"\$?\s*([0-9]{1,3}(?:\.[0-9]+)?)", line)
                if not re.fullmatch(r"20\d{2}", x)
            ]
            if raw_nums:
                out["adj_eps"] = float(raw_nums[-1])
        elif "diluted weighted-average shares outstanding" in low and nums:
            out["shares_diluted"] = float(nums[0])
    return out


def _parse_anf_adjusted_metrics_from_lines(
    lines: List[str],
    *,
    quarter_end: dt.date,
    scale: float,
    source_doc: str,
    source: str,
) -> List[Dict[str, Any]]:
    if not lines or quarter_end is None:
        return []
    statement_values = _parse_anf_statement_values_from_lines(lines, scale=scale)
    quarter_ebitda = _parse_anf_reconciliation_block_values(lines, duration_re=r"\bthirteen\s+weeks?\s+ended\b", scale=scale)
    quarter_ng = _parse_anf_non_gaap_schedule_values(lines, duration_re=r"\bthirteen\s+weeks?\s+ended\b", scale=scale)
    annual_ebitda = _parse_anf_reconciliation_block_values(lines, duration_re=r"\bfifty[- ](?:two|three)\s+weeks?\s+ended\b", scale=scale)
    annual_ng = _parse_anf_non_gaap_schedule_values(lines, duration_re=r"\bfifty[- ](?:two|three)\s+weeks?\s+ended\b", scale=scale)

    rows: List[Dict[str, Any]] = []

    def _base_row(period_type: str, snippet: str) -> Dict[str, Any]:
        return {
            "quarter": quarter_end,
            "period_type": period_type,
            "source_period_label": "FY" if period_type == "annual" else "Q",
            "source": source,
            "source_type": "earnings_financial_schedule",
            "accn": None,
            "filed": None,
            "doc": source_doc,
            "page": None,
            "confidence": "high",
            "col": "ANF financial schedule",
            "source_snippet": snippet,
            "score": 1000,
        }

    q_adj_ebitda = quarter_ebitda.get("adj_ebitda")
    if q_adj_ebitda is None:
        q_adj_ebitda = quarter_ebitda.get("ebitda")
    q_row = _base_row("quarter", "ANF quarterly EBITDA / adjusted EBITDA from earnings financial schedule")
    q_row.update(
        {
            "adj_ebit": quarter_ng.get("adj_ebit", statement_values.get("op_income")),
            "adj_ebitda": q_adj_ebitda,
            "adj_eps": quarter_ng.get("adj_eps", statement_values.get("eps_diluted")),
            "adj_fcf": pd.NA,
        }
    )
    rev_ref = statement_values.get("revenue")
    if rev_ref is not None and pd.notna(rev_ref) and float(rev_ref) > 0:
        for metric_col in ("adj_ebit", "adj_ebitda"):
            metric_val = q_row.get(metric_col)
            if metric_val is not None and pd.notna(metric_val) and float(metric_val) > float(rev_ref) * 5.0:
                q_row[metric_col] = float(metric_val) / 1000.0
    if any(pd.notna(q_row.get(col)) for col in LOCAL_NON_GAAP_CANONICAL_METRICS):
        rows.append(q_row)

    if annual_ebitda or annual_ng:
        ann_adj_ebitda = annual_ebitda.get("adj_ebitda", annual_ebitda.get("ebitda"))
        ann_snippet = "ANF FY adjusted metrics from financial schedule"
        if annual_ng.get("litigation_settlement_adjustment") is not None or annual_ebitda.get("litigation_settlement_adjustment") is not None:
            ann_snippet += "; favorable settlement removed from adjusted results"
        ann_row = _base_row("annual", ann_snippet)
        ann_row.update(
            {
                "adj_ebit": annual_ng.get("adj_ebit"),
                "adj_ebitda": ann_adj_ebitda,
                "adj_eps": annual_ng.get("adj_eps"),
                "adj_fcf": pd.NA,
            }
        )
        if any(pd.notna(ann_row.get(col)) for col in LOCAL_NON_GAAP_CANONICAL_METRICS):
            rows.append(ann_row)

    return rows


def _parse_anf_guidance_rows_from_lines(lines: List[str], q_end: Optional[dt.date]) -> List[Dict[str, Any]]:
    if not lines:
        return []
    norm_lines = [re.sub(r"\s+", " ", str(ln or "")).strip() for ln in lines if str(ln or "").strip()]
    blob = "\n".join(norm_lines)
    low_blob = blob.lower()
    if not re.search(r"\b(fiscal|fy)\s*20\d{2}\b", low_blob, re.I) or not re.search(r"\boutlook|guidance|expects?\b", low_blob, re.I):
        return []
    rows: List[Dict[str, Any]] = []
    seen: set[Tuple[str, str, str]] = set()

    def _infer_guidance_fiscal_year() -> Optional[int]:
        for line in norm_lines:
            line_l = line.lower()
            if "outlook" not in line_l and "guidance" not in line_l and "expects" not in line_l:
                continue
            m_year = re.search(r"\b(?:fiscal|fy)\s*(20\d{2})\b", line, re.I)
            if m_year:
                return int(m_year.group(1))
        m_any = re.search(r"\b(?:fiscal|fy)\s*(20\d{2})\b", blob, re.I)
        if m_any:
            return int(m_any.group(1))
        return q_end.year if isinstance(q_end, dt.date) else None

    fiscal_year = _infer_guidance_fiscal_year()
    quarter_words = {
        "first": 1,
        "second": 2,
        "third": 3,
        "fourth": 4,
    }

    def _infer_outlook_quarter() -> Optional[int]:
        for line in norm_lines:
            line_l = line.lower()
            if "outlook" not in line_l and "guidance" not in line_l:
                continue
            for word, q_num in quarter_words.items():
                if re.search(rf"\b{word}\s+quarter\b", line_l):
                    return q_num
            m_q_line = re.search(r"\bq([1-4])\s*(?:fy|fiscal)?\s*20\d{2}\b", line_l, re.I)
            if m_q_line:
                return int(m_q_line.group(1))
        header_blob = " ".join(norm_lines[:12]).lower()
        for word, q_num in quarter_words.items():
            if re.search(rf"\b{word}\s+quarter\b", header_blob):
                return q_num
        m_q = re.search(r"\bq([1-4])\s*(?:fy|fiscal)?\s*20\d{2}\b", header_blob, re.I)
        if m_q:
            return int(m_q.group(1))
        return None

    q_num = _infer_outlook_quarter()
    quarter_period_label = f"Q{q_num} FY{fiscal_year}" if q_num and fiscal_year else ""
    annual_period_label = f"FY{fiscal_year}" if fiscal_year else ""

    def _number_components(numbers: str) -> Dict[str, Any]:
        txt = str(numbers or "").strip()
        out: Dict[str, Any] = {"low": pd.NA, "high": pd.NA, "value": pd.NA, "unit": ""}
        if not txt:
            return out
        if "bps" in txt.lower():
            vals = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", txt)]
            if len(vals) >= 2:
                out.update({"low": vals[0], "high": vals[1], "unit": "bps"})
            elif vals:
                out.update({"value": vals[0], "unit": "bps"})
            return out
        unit = "$m" if "$" in txt and "million" in txt.lower() else ("%" if "%" in txt else ("m shares" if "share" in txt.lower() or "million" in txt.lower() else ""))
        vals = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", txt)]
        if len(vals) >= 2 and "," in txt:
            out.update({"low": vals[0], "high": vals[1], "unit": unit})
        elif vals:
            out.update({"value": vals[0], "unit": unit})
        return out

    def _add(period_label: str, metric: str, line: str, numbers: str) -> None:
        if not period_label or not metric or not numbers:
            return
        key = (period_label, metric, numbers)
        if key in seen:
            return
        seen.add(key)
        comps = _number_components(numbers)
        rows.append(
            {
                "quarter": q_end,
                "period_label": period_label,
                "period_type": "quarter" if period_label.startswith("Q") else "annual",
                "line": line[:320],
                "numbers": numbers,
                "metric_hint": metric,
                "low": comps.get("low"),
                "high": comps.get("high"),
                "value": comps.get("value"),
                "unit": comps.get("unit"),
            }
        )

    def _fmt_num(raw: str, unit: str, *, force_dollar: bool = False) -> str:
        num = str(raw or "").strip()
        if force_dollar and not num.startswith("$"):
            num = "$" + num
        if unit == "%" and not num.endswith("%"):
            num += "%"
        if unit.lower() == "million" and "million" not in num.lower():
            num += " million"
        return num

    phrase_re = re.compile(
        r"(?P<range_prefix>growth\s+in\s+the\s+range\s+of|in\s+the\s+range\s+of)\s+"
        r"(?P<low_dollar>\$)?(?P<low>\d+(?:\.\d+)?)\s*(?P<low_unit>%|million)?\s*(?:to|-|–)\s*"
        r"(?P<high_dollar>\$)?(?P<high>\d+(?:\.\d+)?)\s*(?P<high_unit>%|million)?"
        r"|(?P<approx_prefix>around|approximately|at least|up to|~)\s*"
        r"(?P<approx_dollar>\$)?(?P<approx>\d+(?:\.\d+)?)\s*(?P<approx_unit>%|million)?",
        re.I,
    )

    def _extract_value_phrases(line: str) -> List[str]:
        out: List[str] = []
        for match in phrase_re.finditer(line):
            if match.group("low") and match.group("high"):
                unit = match.group("high_unit") or match.group("low_unit") or ""
                force_dollar = bool(match.group("low_dollar") or match.group("high_dollar"))
                low = _fmt_num(match.group("low"), unit, force_dollar=force_dollar)
                high = _fmt_num(match.group("high"), unit, force_dollar=force_dollar)
                out.append(f"{low}, {high}")
                continue
            if match.group("approx"):
                prefix = str(match.group("approx_prefix") or "").lower()
                unit = match.group("approx_unit") or ""
                force_dollar = bool(match.group("approx_dollar"))
                val = _fmt_num(match.group("approx"), unit, force_dollar=force_dollar)
                out.append(f"{'~' if prefix == '~' else prefix} {val}".strip())
        return out

    def _metric_for_line(line_l: str) -> Optional[str]:
        if "tariff" in line_l:
            return None
        if "net sales" in line_l:
            return "Revenue"
        if "operating margin" in line_l:
            return "Operating margin"
        if "net income per diluted share" in line_l or re.search(r"\beps\b", line_l):
            return "Adj EPS"
        if "share repurchases" in line_l:
            return "Share repurchases"
        if "diluted weighted average shares" in line_l:
            return "Diluted shares"
        if "capital expenditures" in line_l or re.search(r"\bcapex\b", line_l):
            return "Capex"
        return None

    annual_store_parts: Dict[str, str] = {}
    for line in norm_lines:
        ll = line.lower()
        metric = _metric_for_line(ll)
        if metric:
            phrases = _extract_value_phrases(line)
            if len(phrases) >= 2:
                _add(quarter_period_label, metric, f"{quarter_period_label} {metric} {phrases[0]}", phrases[0])
                _add(annual_period_label, metric, f"{annual_period_label} {metric} {phrases[1]}", phrases[1])
            elif len(phrases) == 1:
                target_period = annual_period_label if metric in {"Capex"} else quarter_period_label
                _add(target_period, metric, f"{target_period} {metric} {phrases[0]}", phrases[0])
        if "real estate activity" in ll or "store openings" in ll or "openings" in ll or "remodel" in ll or "right-size" in ll:
            q_store = re.search(r"(?:~|approximately\s+)?(\d+)\s+net\s+store\s+openings", line, re.I)
            if q_store:
                _add(quarter_period_label, "Real estate activity", f"{quarter_period_label} real estate activity approximately {q_store.group(1)} net store openings", f"~{q_store.group(1)} net store openings")
            open_close = re.search(r"(\d+)\s+openings?\s*,?\s*(\d+)\s+closures?", line, re.I)
            if open_close:
                annual_store_parts["open_close"] = f"{open_close.group(1)} openings, {open_close.group(2)} closures"
            remodel = re.search(r"(\d+)\s+remodels?(?:\s+and\s+|\s*/\s*|\s+/\s+)?right-?sizes?", line, re.I)
            if remodel:
                annual_store_parts["remodels"] = f"{remodel.group(1)} remodels/right-sizes"
        if "tariff impact" in ll or "tariff" in ll:
            bps_vals = re.findall(r"(\d+(?:\.\d+)?)\s*basis\s+points?", line, re.I)
            if len(bps_vals) >= 2:
                _add(quarter_period_label, "Tariffs", f"{quarter_period_label} tariff impact approximately {bps_vals[0]} basis points", f"{bps_vals[0]} bps")
                _add(annual_period_label, "Tariffs", f"{annual_period_label} tariff impact approximately {bps_vals[1]} basis points", f"{bps_vals[1]} bps")
            elif len(bps_vals) == 1:
                target_period = annual_period_label if "full year" in ll else quarter_period_label
                _add(target_period, "Tariffs", f"{target_period} tariff impact approximately {bps_vals[0]} basis points", f"{bps_vals[0]} bps")
    if annual_store_parts:
        numbers = "; ".join(x for x in [annual_store_parts.get("open_close"), annual_store_parts.get("remodels")] if x)
        if numbers:
            _add(annual_period_label, "Real estate activity", f"{annual_period_label} real estate activity {numbers}", numbers)
    return rows


def _parse_anf_pct_value(value: Any) -> Optional[float]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)) and pd.notna(value):
        v = float(value)
        return v / 100.0 if abs(v) > 1.5 else v
    txt = str(value or "").strip()
    if not txt:
        return None
    had_pct = "%" in txt
    neg = bool(re.search(r"\(\s*-?\d", txt))
    txt = txt.replace("%", "").replace(",", "").replace("(", "").replace(")", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", txt)
    if not m:
        return None
    try:
        v = float(m.group(0))
    except Exception:
        return None
    if neg and v > 0:
        v *= -1.0
    return v / 100.0 if had_pct or abs(v) > 1.5 else v


def _parse_anf_amount_thousands(value: Any) -> Optional[float]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)) and pd.notna(value):
        v = float(value)
        if 1900 <= abs(v) <= 2100:
            return None
        return v * 1000.0
    txt = str(value or "").strip()
    if not txt or "%" in txt:
        return None
    raw = coerce_number(txt)
    if raw is None:
        return None
    if 1900 <= abs(float(raw)) <= 2100 and re.fullmatch(r"\d{4}", re.sub(r"\D", "", txt)):
        return None
    return float(raw) * 1000.0


def _anf_segment_label(value: Any) -> str:
    txt = re.sub(r"\s+", " ", str(value or "")).strip()
    low = txt.lower()
    if not low:
        return ""
    if "total" in low and ("company" in low or "comp" in low or "sales" in low):
        return "Total Company"
    if "americas" in low:
        return "Americas"
    if "emea" in low:
        return "EMEA"
    if "apac" in low or "asia pacific" in low:
        return "APAC"
    if "hollister" in low:
        return "Hollister"
    if "abercrombie" in low or re.fullmatch(r"a&f", low):
        return "Abercrombie"
    return ""


def _anf_period_end_for_fy_q(
    fiscal_year: Optional[int],
    quarter_num: Optional[int],
    fiscal_periods: Optional[Dict[Tuple[int, int], dt.date]] = None,
) -> Optional[dt.date]:
    if fiscal_year is None:
        return None
    if quarter_num is None:
        quarter_num = 4
    if fiscal_periods and (int(fiscal_year), int(quarter_num)) in fiscal_periods:
        return fiscal_periods[(int(fiscal_year), int(quarter_num))]
    # Conservative fallback used in unit tests and for metadata-only local files.
    fy = int(fiscal_year)
    q = int(quarter_num)
    fallback = {
        1: dt.date(fy, 5, 3),
        2: dt.date(fy, 8, 2),
        3: dt.date(fy, 11, 1),
        4: dt.date(fy + 1, 1, 31),
    }
    return fallback.get(q)


def _parse_anf_sales_mix_tables_from_html(path_in: Path, q_end: Optional[dt.date]) -> pd.DataFrame:
    """Parse ANF earnings-release region/brand tables into normalized segment rows."""
    try:
        tables = pd.read_html(str(path_in), header=None)
    except Exception:
        return pd.DataFrame()
    rows: List[Dict[str, Any]] = []
    doc = str(path_in)
    for table in tables:
        if table is None or table.empty:
            continue
        table_txt = " ".join([*(str(c) for c in table.columns), *table.astype(str).fillna("").values.ravel().tolist()])
        table_low = table_txt.lower()
        if not any(tok in table_low for tok in ("net sales by segment", "net sales by brand family", "comparable sales")):
            continue
        section = ""
        current_period_type = "annual" if "full year" in table_low and "fourth quarter" not in table_low else "quarter"
        for _, raw_row in table.iterrows():
            cells = [str(x if pd.notna(x) else "").strip() for x in raw_row.tolist()]
            row_txt = " ".join(c for c in cells if c).strip()
            row_low = row_txt.lower()
            if not row_txt:
                continue
            if "full year" in row_low:
                current_period_type = "annual"
                continue
            if "fourth quarter" in row_low or re.search(r"\bquarter\s+fiscal\s+20\d{2}\b", row_low):
                current_period_type = "quarter"
                continue
            if "net sales by segment" in row_low:
                section = "region"
                continue
            if "net sales by brand family" in row_low:
                section = "brand"
                continue
            segment = _anf_segment_label(cells[0] if cells else "")
            if not segment:
                continue
            if segment == "Total Company" and not section:
                section = "total"
            amount: Optional[float] = None
            for cell in cells[1:]:
                amount = _parse_anf_amount_thousands(cell)
                if amount is not None and amount > 1_000_000.0:
                    break
            pct_tokens = re.findall(r"\(?-?\d+(?:\.\d+)?\)?\s*%", row_txt)
            pcts = [_parse_anf_pct_value(tok) for tok in pct_tokens]
            pcts = [float(v) for v in pcts if v is not None]
            growth = pcts[0] if pcts else None
            comp = pcts[-1] if pcts else None
            if amount is not None and amount > 0:
                source_period_label = "FY" if current_period_type == "annual" else "Q"
                rows.append(
                    {
                        "quarter": q_end,
                        "segment": segment,
                        "metric": "revenue",
                        "value": float(amount),
                        "unit": "USD",
                        "period_type": current_period_type,
                        "source_period_label": source_period_label,
                        "source_type": "earnings_release_table",
                        "source": "earnings_release",
                        "doc": doc,
                        "source_snippet": row_txt[:320],
                    }
                )
            if growth is not None and segment != "Total Company":
                source_period_label = "FY" if current_period_type == "annual" else "Q"
                rows.append(
                    {
                        "quarter": q_end,
                        "segment": segment,
                        "metric": "net_sales_growth",
                        "value": float(growth),
                        "unit": "%",
                        "period_type": current_period_type,
                        "source_period_label": source_period_label,
                        "source_type": "earnings_release_table",
                        "source": "earnings_release",
                        "doc": doc,
                        "source_snippet": row_txt[:320],
                    }
                )
            if comp is not None:
                source_period_label = "FY" if current_period_type == "annual" else "Q"
                rows.append(
                    {
                        "quarter": q_end,
                        "segment": segment,
                        "metric": "comparable_sales",
                        "value": float(comp),
                        "unit": "%",
                        "period_type": current_period_type,
                        "source_period_label": source_period_label,
                        "source_type": "earnings_release_table",
                        "source": "earnings_release",
                        "doc": doc,
                        "source_snippet": row_txt[:320],
                    }
                )
    return pd.DataFrame(rows)


def _parse_anf_quarterly_history_retail_driver_rows(
    path_in: Path,
    *,
    fiscal_periods: Optional[Dict[Tuple[int, int], dt.date]] = None,
) -> pd.DataFrame:
    """Parse ANF quarterly-history XLSX comparable sales and store-count schedules."""
    if path_in is None or not Path(path_in).exists():
        return pd.DataFrame()
    try:
        from openpyxl import load_workbook
    except Exception:  # pragma: no cover - optional dependency guard
        return pd.DataFrame()
    try:
        wb = load_workbook(str(path_in), data_only=True, read_only=True)
    except Exception:
        return pd.DataFrame()
    rows: List[Dict[str, Any]] = []
    doc = str(path_in)

    def _fy_from_text(value: Any) -> Optional[int]:
        m = re.search(r"\b(20\d{2})\b", str(value or ""))
        return int(m.group(1)) if m else None

    default_fy = _fy_from_text(Path(path_in).name)

    if "Historical Comparable Sales" in wb.sheetnames:
        ws = wb["Historical Comparable Sales"]
        col_meta: Dict[int, Tuple[Optional[int], Optional[int], str, str]] = {}
        last_fy: Optional[int] = None
        for col in range(2, ws.max_column + 1):
            fy_val = _fy_from_text(ws.cell(row=4, column=col).value)
            if fy_val is not None:
                last_fy = fy_val
            fy = fy_val or last_fy or default_fy
            period_raw = str(ws.cell(row=5, column=col).value or "").strip()
            q_match = re.search(r"\bQ([1-4])\b", period_raw, re.I)
            quarter_num = int(q_match.group(1)) if q_match else None
            period_type = "quarter" if quarter_num is not None else "annual"
            if fy is not None:
                col_meta[col] = (fy, quarter_num, period_type, period_raw)
        for row_idx in range(1, ws.max_row + 1):
            label = str(ws.cell(row=row_idx, column=1).value or "").strip()
            if not label:
                continue
            segment = _anf_segment_label(label)
            if not segment:
                continue
            for col, (fy, q_num, period_type, period_raw) in col_meta.items():
                val = _parse_anf_pct_value(ws.cell(row=row_idx, column=col).value)
                if val is None:
                    continue
                qd = _anf_period_end_for_fy_q(fy, q_num or 4, fiscal_periods)
                if qd is None:
                    continue
                rows.append(
                    {
                        "quarter": qd,
                        "segment": segment,
                        "metric": "comparable_sales",
                        "value": float(val),
                        "unit": "%",
                        "period_type": period_type,
                        "source_period_label": period_raw or ("FY" if period_type == "annual" else f"Q{q_num}"),
                        "source_type": "quarterly_history",
                        "source": "slides",
                        "doc": doc,
                        "source_snippet": f"{label} {period_raw}: {val:.1%}",
                    }
                )

    if "Store Count" in wb.sheetnames:
        ws = wb["Store Count"]
        fy = default_fy
        if fy is None:
            for row_idx in range(1, min(ws.max_row, 12) + 1):
                fy = fy or _fy_from_text(ws.cell(row=row_idx, column=1).value)
        q_end = _anf_period_end_for_fy_q(fy, 4, fiscal_periods) if fy is not None else None
        col_segment: Dict[int, str] = {}
        col_region_brand: Dict[int, Tuple[str, str]] = {}
        current_region = ""
        for col in range(2, ws.max_column + 1):
            region_raw = str(ws.cell(row=4, column=col).value or "").strip()
            if region_raw:
                current_region = _anf_segment_label(region_raw)
            brand_raw = str(ws.cell(row=5, column=col).value or "").strip()
            brand = _anf_segment_label(brand_raw)
            if current_region or brand:
                col_region_brand[col] = (current_region, brand)
            header_bits: List[str] = []
            for row_idx in range(1, 8):
                val = str(ws.cell(row=row_idx, column=col).value or "").strip()
                if val:
                    header_bits.append(val)
            header = " ".join(header_bits)
            seg = _anf_segment_label(header) or ("Total Company" if col == 2 else "")
            if seg:
                col_segment[col] = seg
        metric_map = [
            ("new", "new_stores", "annual", False),
            ("permanently closed", "closed_stores", "annual", True),
            ("closed", "closed_stores", "annual", True),
            ("franchise", "franchise_stores", "quarter", False),
            ("total including franchise", "total_stores_including_franchise", "quarter", False),
        ]
        store_section = "company_owned"
        for row_idx in range(1, ws.max_row + 1):
            label = str(ws.cell(row=row_idx, column=1).value or "").strip()
            low = label.lower()
            if not label or q_end is None:
                continue
            if low == "franchise":
                store_section = "franchise"
                continue
            if low == "total":
                store_section = "total"
                continue
            metric: Optional[str] = None
            period_type = "annual"
            abs_value = False
            if ("january" in low or (fy is not None and str(fy + 1) in low and re.search(r"\b31\b", low))) and store_section == "franchise":
                metric = "franchise_stores"
                period_type = "quarter"
            elif ("january" in low or (fy is not None and str(fy + 1) in low and re.search(r"\b31\b", low))) and store_section == "total":
                metric = "total_stores_including_franchise"
                period_type = "quarter"
            elif "february" in low or (fy is not None and f"{fy}-02" in low):
                metric = "store_count_beginning"
                period_type = "annual"
            elif "january" in low or (fy is not None and str(fy + 1) in low and re.search(r"\b31\b", low)):
                metric = "store_count_end"
                period_type = "quarter"
            for token, metric_name, ptype, make_abs in metric_map:
                if token in low:
                    metric = metric_name
                    period_type = ptype
                    abs_value = make_abs
            if metric is None:
                continue
            row_values: Dict[str, float] = {}
            region_values: Dict[str, float] = {}
            brand_total_values: Dict[str, float] = {}
            for col in range(2, ws.max_column + 1):
                raw = ws.cell(row=row_idx, column=col).value
                num = pd.to_numeric(raw, errors="coerce")
                if pd.isna(num):
                    continue
                value = abs(float(num)) if abs_value else float(num)
                region, brand = col_region_brand.get(col, ("", ""))
                if region in {"Americas", "EMEA", "APAC"}:
                    region_values[region] = region_values.get(region, 0.0) + value
                if region == "Total Company" and brand in {"Abercrombie", "Hollister"}:
                    brand_total_values[brand] = value
                fallback_seg = col_segment.get(col, "")
                if not region_values and fallback_seg:
                    row_values[fallback_seg] = value
            if brand_total_values:
                row_values.update(brand_total_values)
                row_values["Total Company"] = sum(brand_total_values.values())
            elif region_values:
                row_values.update(region_values)
                row_values["Total Company"] = sum(region_values.values())

            for seg, value in row_values.items():
                rows.append(
                    {
                        "quarter": q_end,
                        "segment": seg,
                        "metric": metric,
                        "value": value,
                        "unit": "stores",
                        "period_type": period_type,
                        "source_period_label": "FY" if period_type == "annual" else "Q4",
                        "source_type": "quarterly_history",
                        "source": "slides",
                        "doc": doc,
                        "source_snippet": f"{label} {seg}: {value:g}",
                    }
                )
    try:
        wb.close()
    except Exception:
        pass
    return pd.DataFrame(rows)


def _parse_anf_retail_text_driver_rows_from_lines(
    lines: List[str],
    *,
    quarter_end: Optional[dt.date],
    source_doc: str = "",
    source_type: str = "",
) -> pd.DataFrame:
    """Extract high-signal ANF retail/brand drivers from narrative sources."""
    q_end = quarter_end
    if q_end is None:
        return pd.DataFrame()
    rows: List[Dict[str, Any]] = []
    seen: set[Tuple[str, str]] = set()

    def _clean(txt: Any) -> str:
        return re.sub(r"\s+", " ", str(txt or "")).strip()

    def _add(
        metric: str,
        value: Any,
        unit: str,
        note: str,
        *,
        segment: str = "Total Company",
        period_type: str = "quarter",
        driver_group: str = "ANF retail drivers",
        source_period_label: str = "",
    ) -> None:
        if not metric or value is None:
            return
        note_txt = _clean(note)
        key = (metric, note_txt[:180])
        if key in seen:
            return
        seen.add(key)
        rows.append(
            {
                "quarter": q_end,
                "segment": segment,
                "metric": metric,
                "value": value,
                "unit": unit,
                "period_type": period_type,
                "source_period_label": source_period_label or ("FY" if period_type == "annual" else "Q"),
                "driver_group": driver_group,
                "driver": metric.replace("_", " ").title(),
                "note": note_txt,
                "commentary": note_txt,
                "source_excerpt": note_txt,
                "source_snippet": note_txt,
                "source_doc": source_doc,
                "doc": source_doc,
                "source": source_type,
                "source_type": source_type,
            }
        )

    blob = "\n".join(_clean(line) for line in lines if _clean(line))
    for line in [_clean(line) for line in lines if _clean(line)]:
        low = line.lower()
        if re.search(r"\b44%\s+of\s+total\s+sales\s+were\s+digital\b", low) or ("digital" in low and "44%" in low):
            _add("digital_sales_mix", 0.44, "%", line, period_type="annual", driver_group="Digital / omnichannel")
            holl = re.search(r"hollister\s+(?:around\s+)?(\d+(?:\.\d+)?)%", low)
            ab = re.search(r"abercrombie\s+(?:around\s+)?(\d+(?:\.\d+)?)%", low)
            if holl:
                _add("digital_sales_mix", float(holl.group(1)) / 100.0, "%", line, segment="Hollister", period_type="annual", driver_group="Digital / omnichannel")
            if ab:
                _add("digital_sales_mix", float(ab.group(1)) / 100.0, "%", line, segment="Abercrombie", period_type="annual", driver_group="Digital / omnichannel")
        if "1 billion" in low and re.search(r"\b(visits|platforms?)\b", low):
            _add("digital_visits", 1000.0, "m visits", line, period_type="annual", driver_group="Digital / omnichannel")
        if "omnichannel" in low and ("valuable" in low or "customer" in low):
            _add("omnichannel_customer_value", 1.0, "signal", line, period_type="annual", driver_group="Digital / omnichannel")
        if "record fourth quarter" in low and "brand" in low:
            _add("brand_record_q4_sales", 1.0, "signal", line, driver_group="Brand-family momentum")
        if "abercrombie" in low and "returned to growth" in low:
            _add("abercrombie_returned_to_growth", 1.0, "signal", line, segment="Abercrombie", driver_group="Brand-family momentum")
        if "hollister" in low and ("11th consecutive" in low or "consecutive quarter" in low):
            _add("hollister_consecutive_growth", 11.0, "quarters", line, segment="Hollister", driver_group="Brand-family momentum")
        if "new store" in low or "right size" in low or "remodel" in low or "closed" in low:
            m_new = re.search(r"(\d+)\s+new\s+stores?", low)
            m_rs = re.search(r"(\d+)\s+right\s*sizes?", low)
            m_rem = re.search(r"(\d+)\s+remodels?", low)
            m_closed = re.search(r"closed\s+(\d+)\s+stores?", low) or re.search(r"(\d+)\s+closures?", low)
            m_end = re.search(r"ended\s+the\s+year\s+with\s+(\d+)\s+stores?", low)
            m_holl = re.search(r"(\d+)\s+hollister", low)
            m_ab = re.search(r"(\d+)\s+(?:a&f|abercrombie)", low)
            if m_new:
                _add("new_stores", float(m_new.group(1)), "stores", line, period_type="annual", driver_group="Stores / real estate")
            if m_rs:
                _add("right_sized_stores", float(m_rs.group(1)), "stores", line, period_type="annual", driver_group="Stores / real estate")
            if m_rem:
                _add("remodeled_stores", float(m_rem.group(1)), "stores", line, period_type="annual", driver_group="Stores / real estate")
            if m_closed:
                _add("closed_stores", float(m_closed.group(1)), "stores", line, period_type="annual", driver_group="Stores / real estate")
            if m_end:
                _add("store_count_end", float(m_end.group(1)), "stores", line, driver_group="Stores / real estate")
            if m_holl:
                _add("store_count_end", float(m_holl.group(1)), "stores", line, segment="Hollister", driver_group="Stores / real estate")
            if m_ab:
                _add("store_count_end", float(m_ab.group(1)), "stores", line, segment="Abercrombie", driver_group="Stores / real estate")
        if "inventory" in low:
            m_cost = re.search(r"inventory(?:\s+at)?\s+cost\s+(?:was\s+)?up\s+(\d+(?:\.\d+)?)%", low)
            m_tar = re.search(r"(\d+(?:\.\d+)?)\s+points?\s+from\s+tariffs?", low)
            m_unit = re.search(r"units?\s+(?:were\s+)?up\s+(\d+(?:\.\d+)?)%", low)
            m_erp = re.search(r"(\d+(?:\.\d+)?)\s+points?\s+from\s+strategic\s+receipts|(\d+(?:\.\d+)?)\s+points?\s+.*erp", low)
            m_ex = re.search(r"up\s+roughly\s+(\d+(?:\.\d+)?)%\s+excluding\s+erp", low)
            if m_cost:
                _add("inventory_cost_growth", float(m_cost.group(1)) / 100.0, "%", line, period_type="annual", driver_group="Inventory quality")
            if m_tar:
                _add("inventory_cost_tariff_points", float(m_tar.group(1)), "pts", line, period_type="annual", driver_group="Inventory quality")
            if m_unit:
                _add("inventory_unit_growth", float(m_unit.group(1)) / 100.0, "%", line, period_type="annual", driver_group="Inventory quality")
            if m_erp:
                erp_val = m_erp.group(1) or m_erp.group(2)
                _add("inventory_unit_growth_erp_points", float(erp_val), "pts", line, period_type="annual", driver_group="Inventory quality")
            if m_ex:
                _add("inventory_unit_growth_ex_erp", float(m_ex.group(1)) / 100.0, "%", line, period_type="annual", driver_group="Inventory quality")
        if "tariff" in low or "freight" in low or "erp" in low or "aur" in low or "marketing" in low:
            if "q1" in low or "first quarter" in low:
                m_tar = re.search(r"tariff[^.]{0,80}?(\d+(?:\.\d+)?)\s+basis\s+points?", low)
                m_tar_m = re.search(r"tariff[^.]{0,120}?\$(\d+(?:\.\d+)?)\s+million", low)
                m_fr = re.search(r"freight[^.]{0,80}?(\d+(?:\.\d+)?)\s+basis\s+points?", low)
                m_erp_sales = re.search(r"(\d+(?:\.\d+)?)\s+to\s+(\d+(?:\.\d+)?)\s+percentage\s+point\s+sales\s+headwind", low)
                m_erp_margin = re.search(r"over\s+(\d+(?:\.\d+)?)\s+basis\s+points?\s+of\s+operating\s+margin", low)
                m_marketing = re.search(r"marketing[^.]{0,80}?(\d+(?:\.\d+)?)\s+basis\s+points?", low)
                if m_tar:
                    _add("q1_fy2026_tariff_headwind_bps", float(m_tar.group(1)), "bps", line, driver_group="FY2026 margin bridge", source_period_label="Q1 FY2026")
                if m_tar_m:
                    _add("q1_fy2026_tariff_headwind", float(m_tar_m.group(1)), "$m", line, driver_group="FY2026 margin bridge", source_period_label="Q1 FY2026")
                if m_fr:
                    _add("q1_fy2026_freight_tailwind_bps", float(m_fr.group(1)), "bps", line, driver_group="FY2026 margin bridge", source_period_label="Q1 FY2026")
                if m_erp_sales:
                    _add("q1_fy2026_erp_sales_headwind_low", float(m_erp_sales.group(1)), "ppt", line, driver_group="FY2026 margin bridge", source_period_label="Q1 FY2026")
                    _add("q1_fy2026_erp_sales_headwind_high", float(m_erp_sales.group(2)), "ppt", line, driver_group="FY2026 margin bridge", source_period_label="Q1 FY2026")
                if m_erp_margin:
                    _add("q1_fy2026_erp_margin_headwind_bps", float(m_erp_margin.group(1)), "bps", line, driver_group="FY2026 margin bridge", source_period_label="Q1 FY2026")
                if m_marketing:
                    _add("q1_fy2026_marketing_headwind_bps", float(m_marketing.group(1)), "bps", line, driver_group="FY2026 margin bridge", source_period_label="Q1 FY2026")
            if "fiscal 2026" in low or "fy2026" in low or "full year" in low:
                fy_tar_vals = re.findall(r"(\d+(?:\.\d+)?)\s+basis\s+points?", low)
                m_fy_tar_m = re.search(r"\$(\d+(?:\.\d+)?)\s+million\s+incremental", low)
                if fy_tar_vals:
                    fy_tar = fy_tar_vals[-1] if "full year" in low and len(fy_tar_vals) > 1 else fy_tar_vals[0]
                    _add("fy2026_tariff_headwind_bps", float(fy_tar), "bps", line, period_type="annual", driver_group="FY2026 margin bridge", source_period_label="FY2026")
                if m_fy_tar_m:
                    _add("fy2026_tariff_headwind", float(m_fy_tar_m.group(1)), "$m", line, period_type="annual", driver_group="FY2026 margin bridge", source_period_label="FY2026")
            if "aur" in low:
                _add("fy2026_aur_expansion", 1.0, "signal", line, period_type="annual", driver_group="FY2026 margin bridge", source_period_label="FY2026")
            if "mitigation" in low or "supplier negotiation" in low or "selective pricing" in low:
                _add("tariff_mitigation", 1.0, "signal", line, period_type="annual", driver_group="FY2026 margin bridge", source_period_label="FY2026")
        if "repurchases" in low or "repurchased" in low or "buyback" in low:
            if re.search(r"\b(outlook|guidance|expects?|at least|around)\b", low) and not re.search(r"\b(full year share repurchases|returned|for the full year|repurchased\s+\d)", low):
                continue
            cash_vals = re.findall(r"\$(\d+(?:\.\d+)?)\s+million", low)
            share_vals = re.findall(r"(\d+(?:\.\d+)?)\s+million\s+shares", low)
            paired_cash_val: Optional[str] = None
            paired_share_val: Optional[str] = None
            pair_cash_first = re.search(
                r"(?:share\s+repurchases|repurchases|repurchased|buybacks?)[^.]{0,180}?"
                r"\$(\d+(?:\.\d+)?)\s+million[^.]{0,160}?(\d+(?:\.\d+)?)\s+million\s+shares",
                low,
            )
            pair_shares_first = re.search(
                r"(?:share\s+repurchases|repurchases|repurchased|buybacks?)[^.]{0,180}?"
                r"(\d+(?:\.\d+)?)\s+million\s+shares[^.]{0,160}?\$(\d+(?:\.\d+)?)\s+million",
                low,
            )
            if pair_cash_first:
                paired_cash_val = pair_cash_first.group(1)
                paired_share_val = pair_cash_first.group(2)
            elif pair_shares_first:
                paired_share_val = pair_shares_first.group(1)
                paired_cash_val = pair_shares_first.group(2)
            if paired_cash_val is not None:
                cash_val = paired_cash_val
            elif cash_vals and "remaining" in low and len(cash_vals) > 1:
                cash_val = cash_vals[0]
            elif cash_vals and re.search(r"\b(full year|returned|for the full year)\b", low):
                cash_val = cash_vals[0]
            else:
                cash_val = cash_vals[0] if cash_vals else None
            share_val = (
                paired_share_val
                if paired_share_val is not None
                else (share_vals[-1] if (share_vals and re.search(r"\b(full year|for the full year)\b", low)) else (share_vals[0] if share_vals else None))
            )
            m_pct = re.search(r"(\d+(?:\.\d+)?)%\s+of\s+shares", low)
            m_auth = re.search(r"\$(\d+(?:\.\d+)?)\s+million\s+remaining", low)
            if cash_val:
                _add("share_repurchases", float(cash_val), "$m", line, period_type="annual", driver_group="Capital allocation")
            if share_val:
                _add("shares_repurchased", float(share_val), "m shares", line, period_type="annual", driver_group="Capital allocation")
            if cash_val and share_val and float(share_val) > 0:
                _add("average_buyback_price", float(cash_val) / float(share_val), "$/share", line, period_type="annual", driver_group="Capital allocation")
            if m_pct:
                _add("shares_repurchased_opening_share_pct", float(m_pct.group(1)) / 100.0, "%", line, period_type="annual", driver_group="Capital allocation")
            if m_auth:
                _add("remaining_buyback_authorization", float(m_auth.group(1)), "$m", line, period_type="annual", driver_group="Capital allocation")
    if "store" in blob.lower() and "digital halo" in blob.lower():
        _add("store_digital_halo", 1.0, "signal", "Stores create a digital halo, especially in international markets.", period_type="annual", driver_group="Digital / omnichannel")
    return pd.DataFrame(rows)


def _normalize_anf_guidance_rows(rows_in: Optional[pd.DataFrame]) -> pd.DataFrame:
    if rows_in is None or rows_in.empty:
        return pd.DataFrame()
    df = rows_in.copy()
    for col in ("quarter", "period_label", "metric_hint", "numbers", "line", "source", "doc", "unit"):
        if col not in df.columns:
            df[col] = ""
    df["quarter"] = pd.to_datetime(df["quarter"], errors="coerce")
    df = df[df["quarter"].notna()].copy()
    if df.empty:
        return df
    df["period_label"] = df["period_label"].astype(str).str.strip()
    df["metric_hint"] = df["metric_hint"].astype(str).str.strip()
    df["numbers"] = df["numbers"].astype(str).str.strip()
    df["line"] = df["line"].astype(str).str.strip()
    df["source"] = df["source"].astype(str).str.strip()
    df["doc"] = df["doc"].astype(str).str.strip()
    df["unit"] = df["unit"].astype(str).str.strip()

    def _visible_period_label(period_label: str) -> str:
        raw = str(period_label or "").strip()
        q_match = re.fullmatch(r"Q([1-4])\s*FY(20\d{2})", raw, re.I)
        if q_match:
            return f"Q{int(q_match.group(1))} {int(q_match.group(2))}"
        fy_match = re.fullmatch(r"FY(20\d{2})", raw, re.I)
        if fy_match:
            return f"{int(fy_match.group(1))} year"
        return raw

    def _stated_in_label(q_raw: Any) -> str:
        q_ts = pd.to_datetime(q_raw, errors="coerce")
        if pd.isna(q_ts):
            return ""
        qd = pd.Timestamp(q_ts).date()
        fp = _anf_fiscal_period_from_date(qd)
        if fp is None:
            return ""
        return f"Q{fp[1]} {fp[0]}"

    def _maybe_reclassify_period(period: str, metric: str, line: str) -> Tuple[str, str]:
        period_txt = str(period or "").strip()
        metric_low = str(metric or "").strip().lower()
        line_low = str(line or "").lower()
        annual_context = bool(re.search(r"\b(full[- ]year|full year|fiscal\s+20\d{2}|for the year|annual outlook)\b", line_low, re.I))
        q_match = re.fullmatch(r"Q([1-4])\s+FY(20\d{2})", period_txt, re.I)
        if q_match and metric_low in {"adj eps", "eps"}:
            eps_vals = [
                float(x)
                for x in re.findall(r"\$\s*(\d+(?:\.\d+)?)", str(line or ""))
                if pd.notna(pd.to_numeric(x, errors="coerce"))
            ]
            if any(v >= 5.0 for v in eps_vals):
                return f"FY{int(q_match.group(2))}", "annual"
        if q_match and annual_context and metric_low in {"revenue", "operating margin", "adj eps", "share repurchases", "diluted shares", "capex", "real estate activity", "tariffs"}:
            return f"FY{int(q_match.group(2))}", "annual"
        if period_txt.upper().startswith("FY"):
            return period_txt, "annual"
        if period_txt.upper().startswith("Q"):
            return period_txt, "quarter"
        return period_txt, ""

    def _metric_unit_ok(metric: str, unit: str, numbers: str, line: str) -> bool:
        metric_low = str(metric or "").strip().lower()
        unit_low = str(unit or "").strip().lower()
        blob = f"{numbers} {line}".lower()
        if "revenue" in metric_low:
            return unit_low == "%" or "%" in blob
        if "eps" in metric_low:
            return unit_low in {"", "$", "$/share"} or "$" in blob
        if "operating margin" in metric_low or "margin" == metric_low:
            return unit_low in {"", "%"} or "%" in blob
        if "share repurchase" in metric_low:
            return unit_low in {"", "$m", "$"} and "share" not in unit_low
        if "diluted shares" in metric_low:
            return unit_low in {"", "m shares", "shares"} and "$" not in blob
        if "capex" in metric_low:
            return unit_low in {"", "$m", "$"} or "$" in blob
        if "tariff" in metric_low:
            return unit_low in {"bps", "$m", "$", ""}
        return True

    def _is_business_update_row(source: str, doc: str, line: str) -> bool:
        blob = f"{source} {doc} {line}".lower().replace("-", " ")
        return "business update" in blob or "business_update" in blob or "businessupdate" in blob

    def _is_clean_business_update_same_year(metric: str, line: str, numbers: str) -> bool:
        metric_low = str(metric or "").strip().lower()
        blob = f"{line} {numbers}".lower()
        if "business update" in blob or "currently expects" in blob:
            return True
        if "revenue" in metric_low:
            return "at least" in blob and "6" in blob
        if "operating margin" in metric_low:
            return "around 13" in blob
        if "eps" in metric_low:
            return "10.30" in blob and "10.40" in blob
        if "share repurchase" in metric_low:
            return "450" in blob
        if "capex" in metric_low:
            return "245" in blob
        return False

    def _target_quarter(period: str) -> Optional[Tuple[int, int]]:
        q_match = re.fullmatch(r"Q([1-4])\s+FY(20\d{2})", str(period or "").strip(), re.I)
        if not q_match:
            return None
        return int(q_match.group(2)), int(q_match.group(1))

    allowed_source_re = re.compile(r"\b(earnings_release|press_release|slides|presentation)\b", re.I)
    explicit_context_re = re.compile(
        r"\b(outlook|guidance|expects?|anticipates?|currently expects|business update|fiscal outlook|"
        r"first quarter outlook|full year outlook|in the range|around|at least)\b",
        re.I,
    )
    bad_line_re = re.compile(
        r"\b(actual results?|reported results?|net sales increased|delivered record|table of contents|safe harbor|"
        r"forward-looking statements?|risk factors|annual report|form 10-k)\b",
        re.I,
    )

    keep: List[bool] = []
    for rec in df.to_dict("records"):
        period = str(rec.get("period_label") or "")
        metric = str(rec.get("metric_hint") or "")
        nums = str(rec.get("numbers") or "")
        line = str(rec.get("line") or "")
        source = str(rec.get("source") or "")
        doc = str(rec.get("doc") or "")
        unit = str(rec.get("unit") or "")
        low_line = line.lower()
        if not period or not metric or not nums:
            keep.append(False)
            continue
        period, inferred_period_type = _maybe_reclassify_period(period, metric, line)
        df.loc[df.index[len(keep)], "period_label"] = period
        if inferred_period_type:
            df.loc[df.index[len(keep)], "period_type"] = inferred_period_type
        if not re.match(r"^(?:Q[1-4]\s+FY20\d{2}|FY20\d{2})$", period):
            keep.append(False)
            continue
        if not _metric_unit_ok(metric, unit, nums, line):
            keep.append(False)
            continue
        if not unit and re.search(r"\b(?:approximately|around|~)\s*\d+(?:\.\d+)?\b", nums or line, re.I) and not re.search(r"[%$]|bps|basis points?|shares?|stores?|openings?|closures?", nums or line, re.I):
            keep.append(False)
            continue
        if metric.lower() in {"other", "reported results", "actuals"}:
            keep.append(False)
            continue
        source_ok = (
            bool(allowed_source_re.search(source) or allowed_source_re.search(doc) or "businessupdate" in doc.lower())
            or (not source.strip() and not doc.strip() and bool(explicit_context_re.search(line)))
        )
        if not source_ok:
            keep.append(False)
            continue
        if re.search(r"\.(?:htm|html|pdf|txt|xlsx)\b", low_line) and len(line.split()) <= 4:
            keep.append(False)
            continue
        if bad_line_re.search(line) and not explicit_context_re.search(line):
            keep.append(False)
            continue
        if metric == "Revenue" and ("$" in nums or unit.lower() in {"$m", "usd", "dollars"}):
            keep.append(False)
            continue
        target_year_match = re.search(r"\bFY(20\d{2})\b", period)
        target_fy = int(target_year_match.group(1)) if target_year_match else None
        source_q = pd.to_datetime(rec.get("quarter"), errors="coerce")
        source_period = _anf_fiscal_period_from_date(pd.Timestamp(source_q).date()) if pd.notna(source_q) else None
        source_fy = source_period[0] if source_period is not None else None
        if target_fy is not None and source_fy is not None and (target_fy < source_fy or target_fy > source_fy + 2):
            keep.append(False)
            continue
        target_quarter = _target_quarter(period)
        if target_quarter is not None and source_period is not None:
            target_q_fy, target_q_num = target_quarter
            source_q_num = source_period[1]
            if target_q_fy < source_fy or (target_q_fy == source_fy and target_q_num < source_q_num):
                keep.append(False)
                continue
        if (
            target_fy is not None
            and source_period is not None
            and str(period).upper().startswith("FY")
            and target_fy == source_fy
            and source_period[1] >= 4
        ):
            is_business_update = _is_business_update_row(source, doc, line)
            if not is_business_update or not _is_clean_business_update_same_year(metric, line, nums):
                keep.append(False)
                continue
        keep.append(True)
    out = df.loc[keep].copy()
    if out.empty:
        return out
    if "period_type" not in out.columns:
        out["period_type"] = out["period_label"].str.startswith("Q").map({True: "quarter", False: "annual"})
    else:
        out["period_type"] = out["period_type"].where(out["period_type"].astype(str).str.strip().ne(""), out["period_label"].str.startswith("Q").map({True: "quarter", False: "annual"}))
    if "source_context" not in out.columns:
        out["source_context"] = "normalized_outlook"
    out["horizon_label"] = out["period_label"].map(_visible_period_label)
    out["horizon_type"] = out["period_type"].astype(str).str.lower().map(lambda x: "annual" if x == "annual" else ("quarter" if x == "quarter" else x))
    out["stated_in_label"] = out["quarter"].map(_stated_in_label)
    business_update_mask = out.apply(
        lambda rec: _is_business_update_row(str(rec.get("source") or ""), str(rec.get("doc") or ""), str(rec.get("line") or ""))
        and str(rec.get("horizon_label") or "").strip() == "2025 year",
        axis=1,
    )
    out.loc[business_update_mask, "stated_in_label"] = "Jan 2026 pre-release update"
    out["period_label"] = out["horizon_label"].map(lambda x: re.sub(r"^(20\d{2}) year$", r"FY\1", str(x)) if re.fullmatch(r"20\d{2} year", str(x)) else re.sub(r"^Q([1-4]) (20\d{2})$", r"Q\1 FY\2", str(x)))
    out["_dedupe_key"] = (
        out["quarter"].dt.strftime("%Y-%m-%d")
        + "|"
        + out["period_label"].astype(str)
        + "|"
        + out["metric_hint"].astype(str)
        + "|"
        + out["numbers"].astype(str)
        + "|"
        + out["doc"].astype(str)
    )
    out = out.drop_duplicates("_dedupe_key", keep="first").drop(columns=["_dedupe_key"], errors="ignore")
    return out.reset_index(drop=True)


def _build_anf_guidance_progress_rows(
    slides_guidance: Optional[pd.DataFrame],
    *,
    hist: Optional[pd.DataFrame] = None,
    adj_metrics: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Build a compact ANF guidance-revision progress table from normalized guidance rows."""
    if slides_guidance is None or slides_guidance.empty:
        return pd.DataFrame()
    sg = _normalize_anf_guidance_rows(slides_guidance)
    if sg.empty:
        return pd.DataFrame()
    sg["quarter"] = pd.to_datetime(sg["quarter"], errors="coerce")
    sg = sg[sg["quarter"].notna()].copy()
    sg = sg.sort_values(["period_label", "metric_hint", "quarter"], kind="stable")
    hist_local = hist.copy() if isinstance(hist, pd.DataFrame) and not hist.empty else pd.DataFrame()
    adj_local = adj_metrics.copy() if isinstance(adj_metrics, pd.DataFrame) and not adj_metrics.empty else pd.DataFrame()
    if not hist_local.empty and "quarter" in hist_local.columns:
        hist_local["quarter"] = pd.to_datetime(hist_local["quarter"], errors="coerce")
    if not adj_local.empty and "quarter" in adj_local.columns:
        adj_local["quarter"] = pd.to_datetime(adj_local["quarter"], errors="coerce")

    def _period_complete(period_label: str) -> bool:
        m = re.fullmatch(r"FY(20\d{2})", str(period_label or "").strip())
        if not m or hist_local.empty:
            return False
        fy = int(m.group(1))
        expected_q4 = dt.date(fy + 1, 1, 31)
        return bool((hist_local["quarter"].dt.date == expected_q4).any())

    def _actual_display(period_label: str, metric: str) -> str:
        if hist_local.empty:
            return ""
        m = re.fullmatch(r"FY(20\d{2})", str(period_label or "").strip())
        if not m:
            return ""
        fy = int(m.group(1))
        q4_ts = pd.Timestamp(dt.date(fy + 1, 1, 31))
        cur = hist_local[hist_local["quarter"].eq(q4_ts)].copy()
        prev = hist_local[hist_local["quarter"].eq(pd.Timestamp(dt.date(fy, 2, 1)))].copy()
        if cur.empty:
            return ""
        rec = cur.iloc[-1]
        metric_l = str(metric or "").lower()
        if metric_l == "revenue":
            rev = pd.to_numeric(rec.get("revenue"), errors="coerce")
            if pd.notna(rev) and not prev.empty:
                prev_rev = pd.to_numeric(prev.iloc[-1].get("revenue"), errors="coerce")
                if pd.notna(prev_rev) and abs(float(prev_rev)) > 1e-9:
                    return f"actual +{((float(rev) - float(prev_rev)) / abs(float(prev_rev))) * 100.0:.1f}% sales growth"
            return f"actual ${float(rev) / 1e6:.1f}m revenue" if pd.notna(rev) else ""
        if metric_l == "operating margin":
            rev = pd.to_numeric(rec.get("revenue"), errors="coerce")
            op = pd.to_numeric(rec.get("op_income"), errors="coerce")
            if pd.notna(rev) and pd.notna(op) and abs(float(rev)) > 1e-9:
                return f"actual {float(op) / float(rev) * 100.0:.1f}% GAAP operating margin"
        if metric_l == "adj eps":
            if not adj_local.empty:
                adj = adj_local[adj_local["quarter"].eq(q4_ts)].copy()
                if not adj.empty and "period_type" in adj.columns:
                    adj = adj[adj["period_type"].astype(str).str.lower().eq("annual")]
                if not adj.empty:
                    eps = pd.to_numeric(adj.iloc[-1].get("adj_eps"), errors="coerce")
                    if pd.notna(eps):
                        return f"actual ${float(eps):.2f} adjusted EPS"
            eps = pd.to_numeric(rec.get("eps_diluted"), errors="coerce")
            return f"actual ${float(eps):.2f} GAAP EPS" if pd.notna(eps) else ""
        if metric_l == "capex":
            capex = pd.to_numeric(rec.get("capex"), errors="coerce")
            return f"actual ${float(capex) / 1e6:.1f}m capex" if pd.notna(capex) else ""
        return ""

    metric_display = {
        "Revenue": "Revenue guidance",
        "Operating margin": "Operating margin guidance",
        "Adj EPS": "EPS guidance",
        "Share repurchases": "Share repurchase guidance",
        "Diluted shares": "Diluted-share guidance",
        "Capex": "Capex guidance",
        "Real estate activity": "Real-estate guidance",
        "Tariffs": "Tariff-impact guidance",
    }
    rows: List[Dict[str, Any]] = []
    for (period_label, metric), sub in sg.groupby(["period_label", "metric_hint"], sort=True):
        sub = sub.sort_values("quarter", kind="stable").reset_index(drop=True)
        complete = _period_complete(str(period_label))
        actual_txt = _actual_display(str(period_label), str(metric)) if complete else ""
        for idx, rec in sub.iterrows():
            qv = pd.Timestamp(rec.get("quarter"))
            target = str(rec.get("numbers") or "").strip()
            status = "open"
            rationale = f"{period_label} {metric} guidance: {target}."
            if complete:
                status = "resolved_pass" if actual_txt else "resolved_watch"
                rationale = f"{period_label} {metric} guidance revised to {target}; {actual_txt or 'period complete, actual needs review'}."
            rows.append(
                {
                    "promise_id": f"ANF:{period_label}:{metric}:{idx + 1}",
                    "quarter": qv,
                    "status": status,
                    "score": 90 if complete else 80,
                    "metric_ref": metric_display.get(str(metric), f"{metric} guidance"),
                    "metric_display": metric_display.get(str(metric), f"{metric} guidance"),
                    "target": target,
                    "target_display": target,
                    "actual": actual_txt,
                    "rationale": rationale,
                    "guidance_type": "ANF normalized guidance",
                    "target_period_norm": period_label,
                    "target_period_label": period_label,
                    "first_seen_quarter": qv,
                    "last_seen_quarter": qv,
                    "first_seen_evidence_quarter": qv,
                    "last_seen_evidence_quarter": qv,
                    "source_evidence_json": json.dumps(
                        {
                            "doc": rec.get("doc"),
                            "line": rec.get("line"),
                            "numbers": target,
                            "period_label": period_label,
                            "metric": metric,
                        },
                        ensure_ascii=False,
                        default=str,
                    ),
                    "promise_type": "guidance",
                    "scorable": bool(complete),
                    "numeric_update_this_quarter": True,
                    "evaluated_through": qv,
                }
            )
    return pd.DataFrame(rows)


def _apply_anf_company_overview_overrides(
    overview: Optional[Dict[str, Any]],
    *,
    slides_segments: Optional[pd.DataFrame] = None,
) -> Dict[str, Any]:
    out = dict(overview or {})
    source = "Source: ANF profile fallback / local financial schedules"
    business_fallback = (
        "Abercrombie & Fitch Co. is a global, digitally led omnichannel specialty apparel retailer "
        "operating the Abercrombie and Hollister brand families across stores and digital channels."
    )
    context_fallback = (
        "ANF's current model is driven by brand momentum across Abercrombie and Hollister, comparable sales, "
        "gross margin discipline, inventory control, store optimization, digital/omnichannel engagement, "
        "international growth in EMEA/APAC, and capital returns supported by a net-cash balance sheet. "
        "Digital represented about 44% of FY2025 sales, supported by more than 1 billion platform visits."
    )
    advantage_fallback = (
        "The core advantage is the combination of two refreshed global brand families, an omnichannel store "
        "and digital model, disciplined inventory/markdown management, and balance-sheet flexibility for reinvestment and buybacks."
    )

    noise_re = re.compile(
        r"\b(corporate\s*/\s*other|restricted stock unit|award agreement|code of business conduct|"
        r"governance|proxy statement|exhibit|form of|bylaws?|indemnification|securities act)\b",
        re.I,
    )

    def _bad_text(value: Any) -> bool:
        txt = str(value or "").strip()
        if not txt or txt.upper() == "N/A":
            return True
        return bool(noise_re.search(txt))

    def _weak_anf_context(value: Any) -> bool:
        txt = str(value or "").strip().lower()
        if not txt:
            return True
        if "cash flow and capital allocation" in txt and not re.search(r"\b(abercrombie|hollister|comparable|gross margin|digital|omnichannel|stores?|emea|apac)\b", txt, re.I):
            return True
        return not bool(re.search(r"\b(abercrombie|hollister|comparable|gross margin|inventory|digital|omnichannel|stores?|emea|apac|brand)\b", txt, re.I))

    if _bad_text(out.get("what_it_does")):
        out["what_it_does"] = business_fallback
        out["what_it_does_source"] = source
    if _bad_text(out.get("current_strategic_context")) or _weak_anf_context(out.get("current_strategic_context")):
        out["current_strategic_context"] = context_fallback
        out["current_strategic_context_source"] = source
    if _bad_text(out.get("key_advantage")):
        out["key_advantage"] = advantage_fallback
        out["key_advantage_source"] = source

    revenue_streams: List[Dict[str, Any]] = []
    period_val: Any = None
    if slides_segments is not None and not slides_segments.empty and {"segment", "metric", "value"}.issubset(set(slides_segments.columns)):
        seg = slides_segments.copy()
        seg["segment"] = seg["segment"].astype(str).str.strip()
        seg["metric"] = seg["metric"].astype(str).str.strip().str.lower()
        seg["value"] = pd.to_numeric(seg["value"], errors="coerce")
        if "quarter" in seg.columns:
            seg["quarter"] = pd.to_datetime(seg["quarter"], errors="coerce")
        else:
            seg["quarter"] = pd.NaT
        if "period_type" in seg.columns:
            seg["period_type"] = seg["period_type"].astype(str).str.strip().str.lower()
        else:
            seg["period_type"] = ""
        seg = seg[
            seg["segment"].isin(["Americas", "EMEA", "APAC"])
            & seg["metric"].eq("revenue")
            & seg["value"].notna()
            & (seg["value"] > 0)
        ].copy()
        if not seg.empty:
            annual = seg[seg["period_type"].eq("annual")].copy()
            work = annual if not annual.empty else seg
            best_q = pd.NaT
            best_grp = pd.DataFrame()
            for qv, grp in work.groupby("quarter", dropna=False, sort=True):
                if {"Americas", "EMEA", "APAC"}.issubset(set(grp["segment"])):
                    best_q = qv
                    best_grp = grp
            if not best_grp.empty:
                grouped = best_grp.groupby("segment", as_index=False)["value"].max()
                total = float(grouped["value"].sum())
                if total > 0:
                    order = {"Americas": 0, "EMEA": 1, "APAC": 2}
                    for rec in grouped.sort_values("segment", key=lambda s: s.map(order)).to_dict("records"):
                        revenue_streams.append({"name": rec["segment"], "pct": float(rec["value"]) / total, "value": float(rec["value"])})
                    period_val = best_q if pd.notna(best_q) else None
    if revenue_streams:
        out["revenue_streams"] = revenue_streams
        out["revenue_streams_source"] = "Source: ANF local financial schedules / segment revenue table"
        if period_val is not None:
            out["revenue_streams_period"] = pd.Timestamp(period_val).date()
            out["asof_fy_end"] = pd.Timestamp(period_val).date()

    out["segment_operating_model"] = [
        {"segment": "Americas", "text": "Largest region, with revenue from stores and digital channels across the U.S., Canada and related Americas markets."},
        {"segment": "EMEA", "text": "International growth region served through stores and digital channels, with brand expansion and local market execution."},
        {"segment": "APAC", "text": "Smaller but strategic international region, with APAC revenue separately tracked in ANF segment schedules."},
    ]
    out["segment_operating_model_source"] = source
    out["key_dependencies"] = [
        "Comparable sales and traffic across Abercrombie and Hollister brand families.",
        "Gross margin execution, including product cost, freight, markdowns and tariff mitigation.",
        "Inventory discipline and store/digital omnichannel execution.",
        "International growth in EMEA and APAC.",
        "Buybacks and liquidity supported by cash generation and no core conventional debt in the latest balance sheet.",
    ]
    out["key_dependencies_source"] = source
    return out


def _build_anf_source_quarter_notes(
    *,
    hist: pd.DataFrame,
    base_dir: Path,
    config: PipelineConfig,
    max_quarters: int = 8,
) -> pd.DataFrame:
    if hist is None or hist.empty or "quarter" not in hist.columns:
        return pd.DataFrame()
    h = hist.copy()
    h["quarter"] = pd.to_datetime(h["quarter"], errors="coerce")
    h = h[h["quarter"].notna()].sort_values("quarter")
    target_dates = [pd.Timestamp(q).date() for q in h["quarter"].tail(max_quarters).tolist()]
    if not target_dates:
        return pd.DataFrame()
    target_set = set(target_dates)
    hist_by_q: Dict[dt.date, Dict[str, Any]] = {}
    for rec in h.to_dict("records"):
        qv = pd.to_datetime(rec.get("quarter"), errors="coerce")
        if pd.isna(qv):
            continue
        hist_by_q[pd.Timestamp(qv).date()] = dict(rec)
    fiscal_map = _anf_fiscal_periods_from_history(h)
    aliases = _retail_fiscal_aliases_from_history(h)
    source_dirs: List[Tuple[str, int, Path]] = [
        ("earnings_release", 100, base_dir / "earnings_release"),
        ("presentation", 90, base_dir / "earnings_presentation"),
        ("transcript", 80, base_dir / "earnings_transcripts"),
        ("press_release", 65, base_dir / "press_release"),
        ("transcript", 55, base_dir / "conferences"),
        ("10-k", 20, base_dir / "annual_reports"),
    ]
    theme_patterns: List[Tuple[str, str, str]] = [
        ("Sales", "revenue", r"\b(net sales|sales growth|top line)\b"),
        ("Comparable Sales", "comparable_sales", r"\b(comparable sales|comps?)\b"),
        ("Margin", "margin", r"\b(gross margin|operating margin|product cost|freight|markdown|tariff)\b"),
        ("EPS", "eps", r"\b(eps|diluted share|per diluted share)\b"),
        ("Guidance", "guidance", r"\b(outlook|guidance|expects?|expect|forecast|target)\b"),
        ("Buybacks", "buybacks", r"\b(repurchas|buyback|capital allocation)\b"),
        ("Inventory/Liquidity", "inventory_liquidity", r"\b(inventor|cash|liquidity|balance sheet|marketable securities)\b"),
        ("Brand/Stores/Digital", "operating_drivers", r"\b(abercrombie|hollister|brand|digital|omnichannel|stores?|openings?|closures?|remodels?|right-sizes?|emea|apac)\b"),
    ]
    anti_re = re.compile(
        r"\b(safe harbor|private securities litigation reform act|forward-looking statements?|"
        r"investor contact|media contact|non-gaap financial measures should|not defined or prepared|"
        r"table of contents|exhibit|award agreement|code of business conduct)\b",
        re.I,
    )
    rows: List[Dict[str, Any]] = []
    seen: set[Tuple[dt.date, str, str]] = set()

    def _clean_line(line: Any) -> str:
        txt = re.sub(r"\s+", " ", str(line or "")).strip()
        txt = re.sub(r"\s+([,.;:%])", r"\1", txt)
        return txt

    def _fmt_money_short(val: Any) -> Optional[str]:
        num = pd.to_numeric(val, errors="coerce")
        if pd.isna(num):
            return None
        return f"${float(num) / 1_000_000.0:,.1f}m"

    def _fmt_pct_short(val: Any) -> Optional[str]:
        num = pd.to_numeric(val, errors="coerce")
        if pd.isna(num):
            return None
        return f"{float(num) * 100.0:.1f}%"

    def _close_to_history(metric: str, parsed_val: Any, q_end: dt.date, *, pct_tol: float = 0.08) -> bool:
        hist_val = pd.to_numeric((hist_by_q.get(q_end) or {}).get(metric), errors="coerce")
        parsed_num = pd.to_numeric(parsed_val, errors="coerce")
        if pd.isna(hist_val) or pd.isna(parsed_num):
            return True
        denom = max(abs(float(hist_val)), 1.0)
        return abs(float(parsed_num) - float(hist_val)) <= max(5_000_000.0, denom * pct_tol)

    def _statement_note_values_ok(values: Dict[str, Any], q_end: dt.date) -> bool:
        revenue = pd.to_numeric(values.get("revenue"), errors="coerce")
        gross = pd.to_numeric(values.get("gross_profit"), errors="coerce")
        op_income = pd.to_numeric(values.get("op_income"), errors="coerce")
        if pd.isna(revenue) or pd.isna(gross) or pd.isna(op_income):
            return False
        revenue_f = float(revenue)
        gross_f = float(gross)
        op_f = float(op_income)
        if revenue_f <= 10_000_000.0:
            return False
        gross_margin = gross_f / revenue_f
        op_margin = op_f / revenue_f
        if gross_margin <= 0.0 or gross_margin >= 0.90:
            return False
        if op_margin <= -0.50 or op_margin >= 0.60:
            return False
        return (
            _close_to_history("revenue", revenue_f, q_end)
            and _close_to_history("gross_profit", gross_f, q_end)
            and _close_to_history("op_income", op_f, q_end)
        )

    def _balance_note_values_ok(values: Dict[str, Any], q_end: dt.date) -> bool:
        cash = pd.to_numeric(values.get("cash"), errors="coerce")
        inventory = pd.to_numeric(values.get("inventory"), errors="coerce")
        if pd.isna(cash) or pd.isna(inventory):
            return False
        if float(cash) <= 10_000_000.0 or float(inventory) <= 10_000_000.0:
            return False
        return _close_to_history("cash", cash, q_end, pct_tol=0.10) and _close_to_history("inventory", inventory, q_end, pct_tol=0.10)

    def _append_note(
        *,
        q_end: dt.date,
        category: str,
        metric_ref: str,
        text: str,
        src_type: str,
        path_in: Path,
        priority: int,
        boost: float = 20.0,
    ) -> None:
        txt = _clean_line(text)
        if len(txt) < 24 or anti_re.search(txt) or _looks_low_signal_anf_note(txt):
            return
        key = (q_end, category, re.sub(r"[^a-z0-9]+", " ", txt.lower()).strip()[:140])
        if key in seen:
            return
        seen.add(key)
        note_id = hashlib.sha1(f"ANF|{q_end}|{category}|{key[2]}".encode("utf-8", errors="ignore")).hexdigest()[:12]
        rows.append(
            {
                "quarter": q_end,
                "category": category,
                "tag": category,
                "topic": category,
                "metric_ref": metric_ref,
                "claim": txt,
                "note": txt,
                "source_excerpt": txt,
                "evidence_snippet": txt,
                "evidence_json": json.dumps(
                    {
                        "source_type": src_type,
                        "doc": str(path_in),
                        "quote": txt,
                        "period": str(q_end),
                    },
                    ensure_ascii=False,
                ),
                "source_type": src_type,
                "doc": str(path_in),
                "source_doc": str(path_in),
                "severity": "info",
                "severity_score": float(priority) + float(boost),
                "score": float(priority) + float(boost),
                "rank": 1,
                "note_id": note_id,
            }
        )

    def _guidance_lookup(rows_in: List[Dict[str, Any]]) -> Dict[Tuple[str, str], str]:
        out: Dict[Tuple[str, str], str] = {}
        for rec in rows_in:
            period = str(rec.get("period_label") or "").strip()
            metric = str(rec.get("metric_hint") or "").strip()
            nums = str(rec.get("numbers") or "").strip()
            if period and metric and nums and (period, metric) not in out:
                out[(period, metric)] = nums
        return out

    def _looks_low_signal_anf_note(txt: Any) -> bool:
        t = _clean_line(txt)
        low = t.lower()
        if not t:
            return True
        if re.search(r"\b(form\s+10-k|table of contents|chief executive officer|chief financial officer|vp of investor relations)\b", low):
            return True
        numeric_tokens = re.findall(r"-?\d+(?:,\d{3})*(?:\.\d+)?%?", t)
        if len(numeric_tokens) >= 5 and not re.search(r"\b(growth|expects?|outlook|guidance|inventory|digital|stores?|repurchas\w*|buybacks?|tariff|freight|margin|brand|record|returned|delivered)\b", low):
            return True
        if re.match(r"^(americas|emea|apac|abercrombie|hollister|total company)\b", low) and len(numeric_tokens) >= 3:
            return True
        if low.count(" n/a") >= 2:
            return True
        return False

    for src_type, priority, folder in source_dirs:
        if not folder.exists():
            continue
        folder_files = sorted(folder.glob("*"))
        metadata_primary_raw_targets: set[str] = set()
        if src_type in {"transcript", "conference", "ceo_letter"}:
            for meta_path in folder_files:
                if not meta_path.is_file() or not is_structured_metadata_path(meta_path):
                    continue
                try:
                    meta_values = parse_metadata_key_values(meta_path.read_text(encoding="utf-8", errors="ignore"))
                except Exception:
                    meta_values = {}
                source_file = metadata_source_file(meta_values)
                raw_candidates: List[Path] = []
                if source_file:
                    raw_candidates.append(meta_path.parent / source_file)
                base_stem = re.sub(r"_METADATA_EN$", "", meta_path.stem, flags=re.I)
                if base_stem and base_stem != meta_path.stem:
                    raw_candidates.extend(
                        [
                            meta_path.with_name(f"{base_stem}.txt"),
                            meta_path.with_name(f"{base_stem}.pdf"),
                            meta_path.with_name(f"{base_stem}.htm"),
                            meta_path.with_name(f"{base_stem}.html"),
                        ]
                    )
                for raw_candidate in raw_candidates:
                    try:
                        metadata_primary_raw_targets.add(str(raw_candidate.resolve()).lower())
                    except Exception:
                        metadata_primary_raw_targets.add(str(raw_candidate).lower())
        for path_in in folder_files[:250]:
            if not path_in.is_file() or path_in.suffix.lower() not in {".txt", ".htm", ".html", ".pdf"}:
                continue
            if metadata_primary_raw_targets and not is_structured_metadata_path(path_in):
                try:
                    raw_key = str(path_in.resolve()).lower()
                except Exception:
                    raw_key = str(path_in).lower()
                if raw_key in metadata_primary_raw_targets:
                    continue
            try:
                text, lines = _anf_extract_material_lines(
                    path_in,
                    cache_root=config.cache_dir,
                    rebuild_cache=config.rebuild_doc_text_cache,
                    quiet_pdf_warnings=config.quiet_pdf_warnings,
                )
            except Exception:
                continue
            if not lines:
                continue
            q_end = _infer_anf_quarter_from_material(path_in, text, fiscal_map=fiscal_map, aliases=aliases)
            if q_end not in target_set:
                continue
            scale = _detect_local_non_gaap_text_scale(text)
            if src_type in {"earnings_release", "presentation"}:
                statement_values = _parse_anf_statement_values_from_lines(lines, scale=scale)
                statement_ok = bool(statement_values and _statement_note_values_ok(statement_values, q_end))
                if statement_ok:
                    sales = _fmt_money_short(statement_values.get("revenue"))
                    gross = _fmt_money_short(statement_values.get("gross_profit"))
                    op_inc = _fmt_money_short(statement_values.get("op_income"))
                    gross_margin = _fmt_pct_short(
                        float(statement_values["gross_profit"]) / float(statement_values["revenue"])
                        if statement_values.get("gross_profit") is not None and statement_values.get("revenue")
                        else None
                    )
                    op_margin = _fmt_pct_short(
                        float(statement_values["op_income"]) / float(statement_values["revenue"])
                        if statement_values.get("op_income") is not None and statement_values.get("revenue")
                        else None
                    )
                    if sales and gross and op_inc:
                        _append_note(
                            q_end=q_end,
                            category="Sales / margin",
                            metric_ref="revenue_margin",
                            text=(
                                f"For the quarter ended {q_end.isoformat()}, net sales were {sales}, gross profit was {gross} "
                                f"and operating income was {op_inc}, reflecting gross margin {gross_margin or 'n/a'} "
                                f"and operating margin {op_margin or 'n/a'}."
                            ),
                            src_type=src_type,
                            path_in=path_in,
                            priority=priority,
                        )
                    eps = pd.to_numeric(statement_values.get("eps_diluted"), errors="coerce")
                    shares = pd.to_numeric(statement_values.get("shares_diluted"), errors="coerce")
                    ebitda = _fmt_money_short(statement_values.get("ebitda"))
                    if pd.notna(eps) and pd.notna(shares) and float(shares) > 1_000_000.0:
                        share_txt = f"{float(shares) / 1_000_000.0:,.3f}m"
                        _append_note(
                            q_end=q_end,
                            category="EPS / adjusted EBITDA",
                            metric_ref="eps_adj_ebitda",
                            text=(
                                f"Diluted EPS was ${float(eps):.2f} on {share_txt} diluted shares"
                                + (f", while EBITDA/adjusted EBITDA was {ebitda}" if ebitda else "")
                                + ", supporting the quarter-specific valuation actuals."
                            ),
                            src_type=src_type,
                            path_in=path_in,
                            priority=priority,
                        )
                balance_values = _parse_anf_balance_sheet_values_from_lines(lines, scale=scale)
                if balance_values and _balance_note_values_ok(balance_values, q_end):
                    cash = _fmt_money_short(balance_values.get("cash"))
                    securities = _fmt_money_short(balance_values.get("marketable_securities"))
                    inventory = _fmt_money_short(balance_values.get("inventory"))
                    debt_core = pd.to_numeric(balance_values.get("debt_core"), errors="coerce")
                    if cash and inventory:
                        debt_phrase = "with no conventional core debt identified" if pd.notna(debt_core) and abs(float(debt_core)) < 1 else "with core debt reviewed separately"
                        _append_note(
                            q_end=q_end,
                            category="Inventory / liquidity",
                            metric_ref="inventory_liquidity",
                            text=(
                                f"Balance-sheet support shows cash and equivalents of {cash}"
                                + (f" plus marketable securities of {securities}" if securities else "")
                                + f" and inventory of {inventory}, {debt_phrase}, supporting net-cash flexibility."
                            ),
                            src_type=src_type,
                            path_in=path_in,
                            priority=priority,
                        )
                guidance_rows = _parse_anf_guidance_rows_from_lines(lines, q_end)
                guidance_map = _guidance_lookup(guidance_rows)
                if guidance_map:
                    quarter_labels = sorted({period for period, _metric in guidance_map.keys() if str(period).startswith("Q")})
                    annual_labels = sorted({period for period, _metric in guidance_map.keys() if str(period).startswith("FY")})
                    near_q_label = quarter_labels[0] if quarter_labels else ""
                    fy_label = annual_labels[0] if annual_labels else ""
                    q_rev = guidance_map.get((near_q_label, "Revenue"))
                    q_margin = guidance_map.get((near_q_label, "Operating margin"))
                    q_eps = guidance_map.get((near_q_label, "Adj EPS"))
                    fy_rev = guidance_map.get((fy_label, "Revenue"))
                    fy_margin = guidance_map.get((fy_label, "Operating margin"))
                    fy_eps = guidance_map.get((fy_label, "Adj EPS"))
                    if near_q_label and fy_label and (q_rev or q_margin or q_eps) and (fy_rev or fy_margin or fy_eps):
                        _append_note(
                            q_end=q_end,
                            category="Guidance / outlook",
                            metric_ref="guidance",
                            text=(
                                "Management outlook calls for "
                                f"{near_q_label} net sales growth {q_rev or 'n/a'}, operating margin {q_margin or 'n/a'} and EPS {q_eps or 'n/a'}, "
                                f"with {fy_label} net sales growth {fy_rev or 'n/a'}, operating margin {fy_margin or 'n/a'} and EPS {fy_eps or 'n/a'}."
                            ),
                            src_type=src_type,
                            path_in=path_in,
                            priority=priority,
                        )
                    q_buy = guidance_map.get((near_q_label, "Share repurchases"))
                    fy_buy = guidance_map.get((fy_label, "Share repurchases"))
                    fy_capex = guidance_map.get((fy_label, "Capex"))
                    fy_stores = guidance_map.get((fy_label, "Real estate activity"))
                    q_tariff = guidance_map.get((near_q_label, "Tariffs"))
                    fy_tariff = guidance_map.get((fy_label, "Tariffs"))
                    if near_q_label and fy_label and (q_buy or fy_buy or fy_capex or fy_stores):
                        _append_note(
                            q_end=q_end,
                            category="Capital allocation / stores",
                            metric_ref="buybacks_capex_stores",
                            text=(
                                "The outlook includes "
                                f"{near_q_label} share repurchases {q_buy or 'n/a'}, {fy_label} repurchases {fy_buy or 'n/a'}, "
                                f"capex {fy_capex or 'n/a'} and real-estate activity {fy_stores or 'n/a'}, "
                                f"with tariff impact {q_tariff or 'n/a'} in {near_q_label} and {fy_tariff or 'n/a'} for {fy_label}."
                            ),
                            src_type=src_type,
                            path_in=path_in,
                            priority=priority,
                        )
            retail_rows = _parse_anf_retail_text_driver_rows_from_lines(
                lines,
                quarter_end=q_end,
                source_doc=str(path_in),
                source_type=src_type,
            )
            if retail_rows is not None and not retail_rows.empty:
                metric_values: Dict[str, float] = {}
                try:
                    for retail_rec in retail_rows.to_dict("records"):
                        metric_name = str(retail_rec.get("metric") or "")
                        val_num = pd.to_numeric(retail_rec.get("value"), errors="coerce")
                        if metric_name and pd.notna(val_num) and metric_name not in metric_values:
                            metric_values[metric_name] = float(val_num)
                except Exception:
                    metric_values = {}
                for rec in retail_rows.to_dict("records"):
                    note_txt = str(rec.get("note") or rec.get("source_snippet") or "").strip()
                    metric = str(rec.get("metric") or "")
                    group = str(rec.get("driver_group") or "")
                    if not note_txt:
                        continue
                    category = "Results / drivers / better vs prior"
                    metric_ref = "operating_drivers"
                    if "guidance" in group.lower() or "margin bridge" in group.lower() or "fy2026" in metric.lower():
                        category = "Guidance / outlook"
                        metric_ref = "guidance_margin_bridge"
                    elif "capital allocation" in group.lower() or "repurchase" in metric.lower() or "buyback" in metric.lower():
                        category = "Debt / liquidity / balance sheet"
                        metric_ref = "buybacks"
                    elif "inventory" in group.lower():
                        category = "Inventory / liquidity"
                        metric_ref = "inventory_quality"
                    _append_note(
                        q_end=q_end,
                        category=category,
                        metric_ref=metric_ref,
                        text=note_txt,
                        src_type=src_type,
                        path_in=path_in,
                        priority=priority,
                        boost=30.0,
                    )
                if {"brand_record_q4_sales", "abercrombie_returned_to_growth", "hollister_consecutive_growth"}.intersection(metric_values):
                    _append_note(
                        q_end=q_end,
                        category="Results / drivers / better vs prior",
                        metric_ref="brand_family_momentum",
                        text=(
                            "Q4 FY2025 brand momentum was explicit in source materials: both brands delivered record Q4 net sales, "
                            "Abercrombie returned to growth and Hollister extended its consecutive-quarter growth streak."
                        ),
                        src_type=src_type,
                        path_in=path_in,
                        priority=priority,
                        boost=220.0,
                    )
                if "digital_sales_mix" in metric_values:
                    digital_pct = metric_values.get("digital_sales_mix", 0.0) * 100.0
                    visits = metric_values.get("digital_visits")
                    _append_note(
                        q_end=q_end,
                        category="Results / drivers / better vs prior",
                        metric_ref="digital_omnichannel",
                        text=(
                            f"Digital/omnichannel is sourced as a material ANF driver: digital was about {digital_pct:.0f}% of FY2025 sales"
                            + (f" and ANF platforms generated more than {visits / 1000.0:.0f} billion visits." if visits else ".")
                        ),
                        src_type=src_type,
                        path_in=path_in,
                        priority=priority,
                        boost=215.0,
                    )
                if "inventory_cost_growth" in metric_values and "inventory_unit_growth" in metric_values:
                    inv_cost_pct = metric_values.get("inventory_cost_growth", 0.0) * 100.0
                    inv_unit_pct = metric_values.get("inventory_unit_growth", 0.0) * 100.0
                    inv_tar = metric_values.get("inventory_cost_tariff_points")
                    inv_erp = metric_values.get("inventory_unit_growth_erp_points")
                    inv_ex = metric_values.get("inventory_unit_growth_ex_erp")
                    _append_note(
                        q_end=q_end,
                        category="Inventory / liquidity",
                        metric_ref="inventory_quality",
                        text=(
                            f"Inventory quality note: year-end inventory cost was up about {inv_cost_pct:.0f}%"
                            + (f", including roughly {inv_tar:.0f} pts from tariffs" if inv_tar is not None else "")
                            + f"; units were up about {inv_unit_pct:.0f}%"
                            + (f", including roughly {inv_erp:.0f} pts of ERP prebuild, or {inv_ex * 100.0:.0f}% ex-ERP." if inv_erp is not None and inv_ex is not None else ".")
                        ),
                        src_type=src_type,
                        path_in=path_in,
                        priority=priority,
                        boost=210.0,
                    )
                if "q1_fy2026_tariff_headwind_bps" in metric_values or "fy2026_tariff_headwind_bps" in metric_values:
                    _append_note(
                        q_end=q_end,
                        category="Guidance / outlook",
                        metric_ref="guidance_margin_bridge",
                        text=(
                            "FY2026 margin bridge is sourced: Q1 includes tariff headwind, freight tailwind, ERP operating disruption, "
                            "higher marketing as a percent of sales and modest AUR/selective pricing mitigation."
                        ),
                        src_type=src_type,
                        path_in=path_in,
                        priority=priority,
                        boost=205.0,
                    )
                if "share_repurchases" in metric_values:
                    rep = metric_values.get("share_repurchases")
                    shares = metric_values.get("shares_repurchased")
                    avg = metric_values.get("average_buyback_price")
                    auth = metric_values.get("remaining_buyback_authorization")
                    if rep is not None and rep < 300.0 and shares is None and auth is None:
                        continue
                    _append_note(
                        q_end=q_end,
                        category="Debt / liquidity / balance sheet",
                        metric_ref="buyback_bridge",
                        text=(
                            f"Buyback bridge: 2025 year repurchases were about ${rep:.0f}m"
                            + (f" for {shares:.1f}m shares at roughly ${avg:.2f} per share" if shares is not None and avg is not None else "")
                            + (f", with about ${auth:.0f}m remaining authorization." if auth is not None else ".")
                        ),
                        src_type=src_type,
                        path_in=path_in,
                        priority=priority,
                        boost=205.0,
                    )
            per_theme_count: Dict[str, int] = {}
            for line in lines:
                txt = _clean_line(line)
                if len(txt) < 24 or len(txt) > 360:
                    continue
                low = txt.lower()
                if anti_re.search(low) or _looks_low_signal_anf_note(txt):
                    continue
                for category, metric_ref, pat in theme_patterns:
                    if per_theme_count.get(category, 0) >= 2:
                        continue
                    if not re.search(pat, low, re.I):
                        continue
                    if category == "Guidance" and not re.search(r"\b(outlook|guidance|expects?|in the range|around|at least)\b", low, re.I):
                        continue
                    key = (q_end, category, glx_key := re.sub(r"[^a-z0-9]+", " ", low).strip()[:120])
                    if key in seen:
                        continue
                    seen.add(key)
                    per_theme_count[category] = per_theme_count.get(category, 0) + 1
                    note_id = hashlib.sha1(f"ANF|{q_end}|{category}|{glx_key}".encode("utf-8", errors="ignore")).hexdigest()[:12]
                    rows.append(
                        {
                            "quarter": q_end,
                            "category": category,
                            "tag": category,
                            "topic": category,
                            "metric_ref": metric_ref,
                            "claim": txt,
                            "note": txt,
                            "source_excerpt": txt,
                            "evidence_snippet": txt,
                            "evidence_json": json.dumps(
                                {
                                    "source_type": src_type,
                                    "doc": str(path_in),
                                    "quote": txt,
                                    "period": str(q_end),
                                },
                                ensure_ascii=False,
                            ),
                            "source_type": src_type,
                            "doc": str(path_in),
                            "source_doc": str(path_in),
                            "severity": "info",
                            "severity_score": float(priority),
                            "score": float(priority),
                            "rank": 1,
                            "note_id": note_id,
                        }
                    )
                    break
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows)
    out["quarter"] = pd.to_datetime(out["quarter"], errors="coerce")
    out = out[out["quarter"].notna()].copy()
    out = out.sort_values(["quarter", "severity_score", "category"], ascending=[True, False, True], kind="stable")
    return out.reset_index(drop=True)


def _infer_anf_quarter_from_material(
    path_in: Path,
    text: str,
    *,
    fiscal_map: Dict[Tuple[int, int], dt.date],
    aliases: Dict[dt.date, dt.date],
) -> Optional[dt.date]:
    name = path_in.name
    m = re.search(r"Q([1-4])[_\s-]*(20\d{2})", name, re.I)
    if m:
        q = int(m.group(1))
        fy = int(m.group(2))
        qd = fiscal_map.get((fy, q))
        if qd is not None:
            return qd
    inferred = infer_quarter_end_from_text(text)
    if inferred is not None:
        qd = pd.Timestamp(inferred).date()
        return aliases.get(qd, qd)
    return None


def _apply_anf_local_earnings_financials(
    hist: pd.DataFrame,
    audit: pd.DataFrame,
    *,
    base_dir: Path,
    config: PipelineConfig,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if hist is None or hist.empty or "quarter" not in hist.columns:
        return hist, audit
    aliases = _retail_fiscal_aliases_from_history(hist)
    fiscal_map = _anf_fiscal_periods_from_history(hist)
    quarter_set = set(fiscal_map.values())
    if not quarter_set:
        return hist, audit

    candidates: List[Tuple[int, Path]] = []
    pres_dir = base_dir / "earnings_presentation"
    rel_dir = base_dir / "earnings_release"
    if pres_dir.exists():
        for path_in in sorted(pres_dir.glob("*financial_schedules.pdf")):
            candidates.append((30, path_in))
    if rel_dir.exists():
        for pattern, priority in (("ANF_Q*_earnings_release.pdf", 20), ("8-K_*_earnings_release.htm", 10)):
            for path_in in sorted(rel_dir.glob(pattern)):
                candidates.append((priority, path_in))

    records: Dict[dt.date, Dict[str, Any]] = {}
    for priority, path_in in candidates:
        text, lines = _anf_extract_material_lines(
            path_in,
            cache_root=config.cache_dir,
            rebuild_cache=config.rebuild_doc_text_cache,
            quiet_pdf_warnings=config.quiet_pdf_warnings,
        )
        if not lines:
            continue
        q_end = _infer_anf_quarter_from_material(path_in, text, fiscal_map=fiscal_map, aliases=aliases)
        if q_end is None or q_end not in quarter_set:
            continue
        scale = _detect_local_non_gaap_text_scale(text)
        source_name = (
            "tier3_ex99_pdf_local_earnings_financials"
            if path_in.suffix.lower() == ".pdf"
            else "tier3_ex99_local_earnings_financials"
        )
        rec = records.setdefault(
            q_end,
            {"priority": -1, "values": {}, "ytd": {}, "source_file": path_in.name, "source": source_name},
        )
        values = _parse_anf_statement_values_from_lines(lines, scale=scale)
        ytd = _parse_anf_cash_flow_ytd_from_lines(lines, scale=scale)
        balance_values = _parse_anf_balance_sheet_values_from_lines(lines, scale=scale)
        if priority >= int(rec.get("priority", -1)):
            rec["priority"] = priority
            rec["source_file"] = path_in.name
            rec["source"] = source_name
            for key, value in values.items():
                rec.setdefault("values", {})[key] = value
            for key, value in balance_values.items():
                rec.setdefault("values", {})[key] = value
            for key, value in ytd.items():
                rec.setdefault("ytd", {})[key] = value
        else:
            for key, value in ytd.items():
                rec.setdefault("ytd", {}).setdefault(key, value)
            for key, value in values.items():
                rec.setdefault("values", {}).setdefault(key, value)
            for key, value in balance_values.items():
                rec.setdefault("values", {}).setdefault(key, value)

    ytd_by_period: Dict[Tuple[int, int], Dict[str, Any]] = {}
    for q_end, rec in records.items():
        fq = _anf_fiscal_period_from_date(q_end)
        if fq is not None and rec.get("ytd"):
            ytd_by_period[fq] = rec
    for (fy, q), rec in sorted(ytd_by_period.items()):
        ytd = rec.get("ytd") or {}
        prev = ytd_by_period.get((fy, q - 1), {}) if q > 1 else {}
        prev_ytd = prev.get("ytd") or {}
        values = rec.setdefault("values", {})
        if "cfo_ytd" in ytd:
            values["cfo"] = float(ytd["cfo_ytd"]) - float(prev_ytd.get("cfo_ytd", 0.0) or 0.0)
        if "capex_ytd" in ytd:
            capex_val = float(ytd["capex_ytd"]) - float(prev_ytd.get("capex_ytd", 0.0) or 0.0)
            if capex_val >= 0:
                values["capex"] = capex_val

    hist_out = hist.copy()
    hist_q = pd.to_datetime(hist_out["quarter"], errors="coerce").dt.date
    audit_rows: List[Dict[str, Any]] = []
    for q_end, rec in sorted(records.items()):
        values = rec.get("values") or {}
        if not values:
            continue
        mask = hist_q == q_end
        if not bool(mask.any()):
            continue
        for metric, value in values.items():
            if metric not in hist_out.columns:
                hist_out[metric] = pd.NA
            current = pd.to_numeric(hist_out.loc[mask, metric], errors="coerce")
            should_set = bool(current.isna().all())
            if metric in {"cfo", "capex", "shares_diluted", "ebitda", "cash", "inventory", "marketable_securities", "debt_core"} and current.notna().any():
                cur_val = float(current.dropna().iloc[0])
                tight_metrics = {"shares_diluted", "cash", "inventory", "marketable_securities", "debt_core"}
                tolerance = max(1.0, abs(float(value)) * (0.002 if metric in tight_metrics else 0.02))
                should_set = abs(cur_val - float(value)) > tolerance
            if not should_set:
                continue
            hist_out.loc[mask, metric] = float(value)
            audit_rows.append(
                {
                    "metric": metric,
                    "quarter": q_end,
                    "source": rec.get("source") or "tier3_ex99_local_earnings_financials",
                    "tag": rec.get("source_file"),
                    "accn": None,
                    "form": "local_material",
                    "filed": None,
                    "start": None,
                    "end": q_end,
                    "unit": "USD",
                    "duration_days": None,
                    "value": float(value),
                    "note": "ANF local earnings financials fallback; cash-flow metrics derived from cumulative schedules when needed",
                }
            )
    if audit_rows:
        audit_add = pd.DataFrame(audit_rows)
        audit = audit_add if audit is None or audit.empty else pd.concat([audit, audit_add], ignore_index=True)
    return hist_out, audit


def _normalized_quarter_timestamps(values: Any) -> set[pd.Timestamp]:
    if values is None:
        return set()
    q_series = pd.to_datetime(values, errors="coerce")
    if not isinstance(q_series, pd.Series):
        q_series = pd.Series(q_series)
    return {pd.Timestamp(v) for v in q_series.dropna()}


def _filter_missing_local_non_gaap_quarters(
    q_targets: List[dt.date],
    existing_q: set[pd.Timestamp],
) -> List[dt.date]:
    out: List[dt.date] = []
    seen: set[pd.Timestamp] = set()
    for qd in q_targets:
        try:
            q_ts = pd.Timestamp(qd)
        except Exception:
            continue
        if q_ts in existing_q or q_ts in seen:
            continue
        seen.add(q_ts)
        out.append(qd)
    return out


def _existing_local_non_gaap_metrics_by_quarter(
    df: Optional[pd.DataFrame],
    *,
    metrics: Tuple[str, ...] = LOCAL_NON_GAAP_CANONICAL_METRICS,
) -> Dict[pd.Timestamp, set[str]]:
    out: Dict[pd.Timestamp, set[str]] = {}
    if df is None or df.empty or "quarter" not in df.columns:
        return out
    q_series = pd.to_datetime(df["quarter"], errors="coerce")
    if not isinstance(q_series, pd.Series):
        q_series = pd.Series(q_series, index=df.index)
    available_metrics = [metric for metric in metrics if metric in df.columns]
    if not available_metrics:
        return out
    for idx, q_val in q_series.items():
        if pd.isna(q_val):
            continue
        present_metrics = {metric for metric in available_metrics if pd.notna(df.at[idx, metric])}
        if present_metrics:
            out.setdefault(pd.Timestamp(q_val), set()).update(present_metrics)
    return out


def _filter_missing_local_non_gaap_metric_quarters(
    q_targets: List[dt.date],
    existing_metrics_by_quarter: Dict[pd.Timestamp, set[str]],
    *,
    metrics: Tuple[str, ...] = LOCAL_NON_GAAP_CANONICAL_METRICS,
) -> List[dt.date]:
    out: List[dt.date] = []
    seen: set[pd.Timestamp] = set()
    wanted_metrics = {metric for metric in metrics if metric}
    for qd in q_targets:
        try:
            q_ts = pd.Timestamp(qd)
        except Exception:
            continue
        if q_ts in seen:
            continue
        seen.add(q_ts)
        if wanted_metrics.difference(existing_metrics_by_quarter.get(q_ts, set())):
            out.append(qd)
    return out


def _prune_local_non_gaap_metrics_against_existing(
    local_metrics: Optional[pd.DataFrame],
    existing_metrics_by_quarter: Dict[pd.Timestamp, set[str]],
    *,
    metrics: Tuple[str, ...] = LOCAL_NON_GAAP_CANONICAL_METRICS,
) -> pd.DataFrame:
    if local_metrics is None or local_metrics.empty:
        return pd.DataFrame() if local_metrics is None else local_metrics
    if "quarter" not in local_metrics.columns:
        return local_metrics

    pruned = local_metrics.copy()
    pruned["quarter"] = pd.to_datetime(pruned["quarter"], errors="coerce")
    pruned = pruned[pruned["quarter"].notna()].copy()
    if pruned.empty:
        return pruned

    tracked_metrics = [metric for metric in metrics if metric in pruned.columns]
    if not tracked_metrics:
        return pruned

    keep_indices: List[int] = []
    for idx, quarter in pruned["quarter"].items():
        existing_metrics = existing_metrics_by_quarter.get(pd.Timestamp(quarter), set())
        for metric in tracked_metrics:
            if metric in existing_metrics:
                pruned.at[idx, metric] = pd.NA
        if any(pd.notna(pruned.at[idx, metric]) for metric in tracked_metrics):
            keep_indices.append(idx)

    if not keep_indices:
        return pruned.iloc[0:0].copy()
    return pruned.loc[keep_indices].reset_index(drop=True)


def _sanitize_anf_adjusted_metric_units(
    adj_metrics: pd.DataFrame,
    adj_breakdown: pd.DataFrame,
    hist: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Normalize ANF adjusted metrics that generic SEC HTML parsing reads 1,000x high."""
    if adj_metrics is None or adj_metrics.empty:
        return adj_metrics, adj_breakdown
    out = adj_metrics.copy()
    if "quarter" not in out.columns:
        return out, adj_breakdown
    out["quarter"] = pd.to_datetime(out["quarter"], errors="coerce")

    revenue_by_q: Dict[pd.Timestamp, float] = {}
    if hist is not None and not hist.empty and {"quarter", "revenue"}.issubset(hist.columns):
        h = hist.copy()
        h["quarter"] = pd.to_datetime(h["quarter"], errors="coerce")
        h["revenue"] = pd.to_numeric(h["revenue"], errors="coerce")
        for rec in h.dropna(subset=["quarter", "revenue"]).to_dict("records"):
            revenue_by_q[pd.Timestamp(rec["quarter"])] = float(rec["revenue"])

    scaled_quarters: set[pd.Timestamp] = set()
    metric_cols = [c for c in ("adj_ebit", "adj_ebitda", "adj_fcf") if c in out.columns]
    for idx, row in out.iterrows():
        q_raw = row.get("quarter")
        if pd.isna(q_raw):
            continue
        q_ts = pd.Timestamp(q_raw)
        rev = revenue_by_q.get(q_ts)
        for col in metric_cols:
            val = pd.to_numeric(row.get(col), errors="coerce")
            if pd.isna(val):
                continue
            val_f = float(val)
            abs_val = abs(val_f)
            scale_down = False
            if abs_val >= 5_000_000_000.0:
                scale_down = True
            elif rev is not None and rev > 0 and abs_val > rev * 10.0 and (abs_val / 1000.0) <= rev * 1.5:
                scale_down = True
            if scale_down:
                out.at[idx, col] = val_f / 1000.0
                scaled_quarters.add(q_ts)
                if "source_snippet" in out.columns:
                    prior_note = str(out.at[idx, "source_snippet"] or "")
                    note = "ANF unit-normalized from 1,000x generic SEC parse"
                    if note not in prior_note:
                        out.at[idx, "source_snippet"] = (prior_note + " | " + note).strip(" |")

    if adj_breakdown is None or adj_breakdown.empty or "quarter" not in adj_breakdown.columns or "value" not in adj_breakdown.columns:
        return out, adj_breakdown
    bd = adj_breakdown.copy()
    bd["quarter"] = pd.to_datetime(bd["quarter"], errors="coerce")
    bd["value"] = pd.to_numeric(bd["value"], errors="coerce")
    for idx, row in bd.iterrows():
        q_raw = row.get("quarter")
        val = row.get("value")
        if pd.isna(q_raw) or pd.isna(val):
            continue
        q_ts = pd.Timestamp(q_raw)
        val_f = float(val)
        rev = revenue_by_q.get(q_ts)
        if abs(val_f) >= 5_000_000_000.0 or (
            q_ts in scaled_quarters
            and rev is not None
            and rev > 0
            and abs(val_f) > rev * 10.0
            and (abs(val_f) / 1000.0) <= rev * 1.5
        ):
            bd.at[idx, "value"] = val_f / 1000.0
    return out, bd


def _local_non_gaap_pdf_cache_dirs(base_dir: Path, src_name: str) -> Tuple[Path, Path]:
    cache_root = preferred_ticker_cache_root_from_base_dir(base_dir)
    if str(src_name or "").strip().lower() == "slides":
        return cache_root / "slides_text", cache_root / "slides_ocr"
    src_key = re.sub(r"[^a-z0-9]+", "_", str(src_name or "other").strip().lower()).strip("_") or "other"
    return cache_root / f"local_non_gaap_{src_key}_text", cache_root / f"local_non_gaap_{src_key}_ocr"


def _local_non_gaap_pdf_cache_key(path_in: Path, *, src_name: str, page_number: int) -> str:
    if str(src_name or "").strip().lower() == "slides":
        return f"{path_in.stem}_p{page_number}"
    try:
        raw_key = str(path_in.resolve())
    except Exception:
        raw_key = str(path_in)
    digest = hashlib.sha1(
        f"{LOCAL_NON_GAAP_PDF_PAGE_CACHE_VERSION}|{str(src_name or '').strip().lower()}|{raw_key}".encode(
            "utf-8",
            errors="ignore",
        )
    ).hexdigest()[:16]
    src_key = re.sub(r"[^a-z0-9]+", "_", str(src_name or "other").strip().lower()).strip("_") or "other"
    return f"{src_key}_{digest}_p{page_number}"


def run_pipeline_impl(
    config: PipelineConfig,
    sec_config: SecConfig,
    *,
    ticker: Optional[str] = None,
    cik: Optional[str] = None,
) -> PipelineArtifacts:
    """Build the full artifact bundle consumed by workbook export.

    This is the main assembly boundary for the runtime. Upstream we have SEC cache,
    local materials, and market/source refresh state; downstream we hand a normalized
    `PipelineArtifacts` bundle to the thin `pipeline.py` surface and eventually into
    `WorkbookInputs`.
    """
    sec = SecClient(cache_dir=config.cache_dir, cfg=sec_config)
    repo_root = (
        Path(config.repo_root).expanduser().resolve()
        if config.repo_root is not None
        else Path(__file__).resolve().parents[2]
    )
    default_base_dir = (
        Path(config.material_root).expanduser().resolve()
        if config.material_root is not None
        else Path(__file__).resolve().parents[1]
    )
    tkr_raw, tkr_u, base_dir = resolve_pipeline_roots(
        repo_root=repo_root,
        default_base_dir=default_base_dir,
        ticker=ticker,
        material_root=config.material_root,
    )

    if cik:
        cik_int = int(cik)
    elif ticker:
        cik_int = cik_from_ticker(sec, ticker)
    else:
        raise RuntimeError("Must provide ticker or cik")

    cik10 = cik10_from_int(cik_int)

    cf = sec.companyfacts(cik10)
    sub = sec.submissions(cik10)
    df_all = companyfacts_to_df(cf, namespace=config.namespace)
    # Defensive: keep fy_calc available even if upstream normalization changes.
    if df_all is not None and not df_all.empty and "fy_calc" not in df_all.columns:
        fy_end_mmdd = (12, 31)
        try:
            if "fp" in df_all.columns and "end_d" in df_all.columns:
                fy_rows = df_all[df_all["fp"].astype(str).str.upper().isin(["FY", "Q4"])].copy()
                fy_rows = fy_rows[fy_rows["end_d"].notna()]
                if not fy_rows.empty:
                    mmdd = fy_rows["end_d"].map(lambda d: (d.month, d.day))
                    if not mmdd.empty:
                        fy_end_mmdd = mmdd.value_counts().idxmax()
        except Exception:
            fy_end_mmdd = (12, 31)
        def _calc_fy_fallback(end_d: Any) -> Optional[int]:
            if end_d is None or pd.isna(end_d):
                return None
            try:
                if (int(end_d.month), int(end_d.day)) > (int(fy_end_mmdd[0]), int(fy_end_mmdd[1])):
                    return int(end_d.year) + 1
                return int(end_d.year)
            except Exception:
                return None
        df_all["fy_calc"] = df_all["end_d"].map(_calc_fy_fallback)
    stage_timings: Dict[str, float] = {}
    local_material_sig = material_dirs_signature(base_dir, ticker)
    # Stage cache is the fine-grained persistence layer. Its role is to reuse expensive
    # intermediate frames while still allowing submissions, code, and material changes
    # to invalidate stale outputs predictably.
    stage_cache = PipelineStageCache(Path(config.cache_dir) / "pipeline_stage_cache", cik10, PIPELINE_STAGE_CACHE_VERSION)
    _sub_recent_signature = submissions_recent_signature
    _df_quick_sig = dataframe_quick_signature
    _load_stage_cache = stage_cache.load
    _save_stage_cache = stage_cache.save
    _timed_stage = timed_stage

    submissions_sig = _sub_recent_signature(sub, forms_prefix=("10-Q", "10-K", "8-K", "DEF 14A", "DEFA14A"), max_rows=600)
    df_all_sig = _df_quick_sig(df_all, ["concept", "end_d", "start_d", "val", "fy_calc", "fp", "frame"])
    # GAAP history is the first expensive stage because many later stages depend on
    # quarter-normalized history, audit rows, and preview tables. Cache invalidation
    # is driven by recent submissions identity plus a compact signature of facts.
    gaap_history_key = "|".join(
        [
            "v2",
            f"sub={submissions_sig}",
            f"facts={df_all_sig}",
            f"max_q={config.max_quarters}",
            f"min_year={config.min_year}",
            f"strict={config.strictness}",
            f"ticker={str(ticker or '').upper()}",
        ]
    )
    gaap_cached = _load_stage_cache("gaap_history_bundle", gaap_history_key)
    if isinstance(gaap_cached, dict):
        hist = gaap_cached.get("hist", pd.DataFrame())
        audit = gaap_cached.get("audit", pd.DataFrame())
        qfd_preview = gaap_cached.get("qfd_preview", pd.DataFrame())
        qfd_unused = gaap_cached.get("qfd_unused", pd.DataFrame())
    else:
        with _timed_stage(stage_timings, "gaap_history", enabled=config.profile_timings):
            hist, audit, qfd_preview, qfd_unused = build_gaap_history(
                df_all,
                max_quarters=config.max_quarters,
                strictness=config.strictness,
                min_year=config.min_year,
                sec=sec,
                cik_int=cik_int,
                submissions=sub,
                ticker=ticker,
                quiet_pdf_warnings=config.quiet_pdf_warnings,
                stage_timings=stage_timings,
                profile_timings=config.profile_timings,
            )
        _save_stage_cache(
            "gaap_history_bundle",
            gaap_history_key,
            {
                "hist": hist,
                "audit": audit,
                "qfd_preview": qfd_preview,
                "qfd_unused": qfd_unused,
            },
        )
    # Imported lazily to avoid widening the pipeline/module dependency cycle.
    from .pipeline import build_tag_coverage

    tag_coverage = build_tag_coverage(df_all)
    if tkr_u == "ANF":
        with _timed_stage(stage_timings, "anf_local_earnings_financials", enabled=config.profile_timings):
            hist, audit = _apply_anf_local_earnings_financials(
                hist,
                audit,
                base_dir=base_dir,
                config=config,
            )
    period_checks = self_check_period_logic(
        df_all,
        audit,
        metric_name="revenue",
        strictness=config.strictness,
    )
    debt_tranches = pd.DataFrame()
    if config.enable_tier2_debt:
        # Tier-2 debt is cached independently because it is expensive, SEC-driven,
        # and reused by both workbook debt tabs and downstream QA.
        debt_tranches_key = "|".join(
            [
                # v2 invalidates stale debt-table parses after summary rows with
                # shifted current-period amounts were aligned to the date header.
                "v2",
                f"sub={submissions_sig}",
                f"max_q={config.max_quarters}",
                f"min_year={config.min_year}",
            ]
        )
        debt_tranches_cached = _load_stage_cache("debt_tranches_tier2", debt_tranches_key)
        if isinstance(debt_tranches_cached, pd.DataFrame):
            debt_tranches = debt_tranches_cached
        else:
            with _timed_stage(stage_timings, "debt_tranches_tier2", enabled=config.profile_timings):
                debt_tranches = build_debt_tranches_tier2(
                    sec,
                    cik_int,
                    sub,
                    max_quarters=config.max_quarters,
                    min_year=config.min_year,
                )
            _save_stage_cache("debt_tranches_tier2", debt_tranches_key, debt_tranches)
    qa_checks = build_qa_checks(df_all, hist, audit=audit)
    debt_qa = build_debt_qa_checks(debt_tranches)
    if debt_qa is not None and not debt_qa.empty:
        if qa_checks is None or qa_checks.empty:
            qa_checks = debt_qa
        else:
            qa_checks = pd.concat([qa_checks, debt_qa], ignore_index=True)
    interest_qa = build_interest_qa_checks(hist, audit)
    if interest_qa is not None and not interest_qa.empty:
        if qa_checks is None or qa_checks.empty:
            qa_checks = interest_qa
        else:
            qa_checks = pd.concat([qa_checks, interest_qa], ignore_index=True)
    debt_buckets = pd.DataFrame()
    debt_bucket_qa = pd.DataFrame()
    bridge_q = build_bridge_q(hist)

    adj_metrics = pd.DataFrame()
    adj_breakdown = pd.DataFrame()
    non_gaap_files = pd.DataFrame()
    adj_metrics_relaxed = pd.DataFrame()
    adj_breakdown_relaxed = pd.DataFrame()
    non_gaap_files_relaxed = pd.DataFrame()
    slides_segments = pd.DataFrame()
    slides_debt = pd.DataFrame()
    slides_guidance = pd.DataFrame()
    slides_guidance_raw = pd.DataFrame()

    if config.enable_tier3_non_gaap:
        def _load_or_build_tier3(mode_name: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
            # Tier-3 non-GAAP outputs are persisted per mode because strict and
            # relaxed runs intentionally have different evidence and suppression
            # rules while feeding the same workbook surfaces.
            tier3_key = "|".join(
                [
                    "v2",
                    f"sub={submissions_sig}",
                    f"mode={mode_name}",
                    f"max_q={config.max_quarters}",
                ]
            )
            cached = _load_stage_cache(f"tier3_non_gaap_{mode_name}", tier3_key)
            if isinstance(cached, dict):
                return (
                    cached.get("metrics", pd.DataFrame()),
                    cached.get("breakdown", pd.DataFrame()),
                    cached.get("files", pd.DataFrame()),
                )
            with _timed_stage(stage_timings, f"tier3_non_gaap_{mode_name}", enabled=config.profile_timings):
                m_df, b_df, f_df = build_non_gaap_tier3(sec, cik_int, sub, max_quarters=config.max_quarters, mode=mode_name)
            _save_stage_cache(
                f"tier3_non_gaap_{mode_name}",
                tier3_key,
                {"metrics": m_df, "breakdown": b_df, "files": f_df},
            )
            return m_df, b_df, f_df

        if config.non_gaap_mode == "relaxed":
            adj_metrics, adj_breakdown, non_gaap_files = _load_or_build_tier3("relaxed")
        else:
            adj_metrics, adj_breakdown, non_gaap_files = _load_or_build_tier3("strict")
            if config.non_gaap_preview:
                adj_metrics_relaxed, adj_breakdown_relaxed, non_gaap_files_relaxed = _load_or_build_tier3("relaxed")

    # Last-resort local fallback for adjusted metrics (slides/transcripts) when EX-99 missing
    local_period_aliases = _retail_fiscal_aliases_from_history(hist) if str(ticker or "").strip().upper() == "ANF" else {}
    anf_fiscal_map = _anf_fiscal_periods_from_history(hist) if str(ticker or "").strip().upper() == "ANF" else {}

    def _resolve_local_period_end(qd: Optional[dt.date]) -> Optional[dt.date]:
        if qd is None:
            return None
        try:
            d0 = pd.Timestamp(qd).date()
        except Exception:
            return qd
        return local_period_aliases.get(d0, d0)

    def _infer_q_from_filename(name: str) -> Optional[dt.date]:
        return _resolve_local_period_end(_infer_local_non_gaap_period_end_from_name(name))

    def _extract_text_from_file(p: Path) -> str:
        suf = p.suffix.lower()
        if suf in (".txt",):
            return p.read_text(encoding="utf-8", errors="ignore")
        if suf in (".htm", ".html"):
            return strip_html(p.read_text(encoding="utf-8", errors="ignore"))
        if suf == ".pdf":
            return extract_pdf_text_cached(
                p,
                cache_root=config.cache_dir,
                rebuild_cache=config.rebuild_doc_text_cache,
                quiet_pdf_warnings=config.quiet_pdf_warnings,
            )
        return ""

    existing_metrics_by_quarter_for_local_fallback: Dict[pd.Timestamp, set[str]] = {}

    def build_non_gaap_local_fallback() -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        rows_m: List[Dict[str, Any]] = []
        rows_f: List[Dict[str, Any]] = []
        rows_seg: List[Dict[str, Any]] = []
        rows_debt: List[Dict[str, Any]] = []
        rows_guid: List[Dict[str, Any]] = []
        has_financial_statement_files = _local_non_gaap_has_financial_statement_files(base_dir, tkr_u)
        sources = [
            ("earnings_release", base_dir / "earnings_release"),
            ("earnings_release", base_dir / "Earnings Release"),
            ("earnings_release", base_dir / "Earnings Releases"),
            ("press_release", base_dir / "press_release"),
            ("press_release", base_dir / "Press Release"),
            ("slides", base_dir / "slides"),
            ("slides", base_dir / "earnings_presentation"),
            ("slides", base_dir / "Earnings Presentation"),
            ("transcripts", base_dir / "Earnings Transcripts"),
            ("transcripts", base_dir / "transcripts"),
            ("transcripts", base_dir / "earnings_transcripts"),
            ("annual_reports", base_dir / "annual_reports"),
            ("financial_statement", base_dir / "financial_statement"),
        ]
        if tkr_u:
            sources.extend(
                [
                    ("financial_statement", base_dir / f"{tkr_u}-10K"),
                    ("financial_statement", base_dir / f"{tkr_u}_10K"),
                    ("financial_statement", base_dir / f"{tkr_u} 10K"),
                ]
            )
        seen_q: set[pd.Timestamp] = set()

        def _allow_actuals_from_local_page(src_name: str, path_in: Path, txt: str) -> bool:
            # Preliminary press releases can update guidance/narrative, but they
            # are not complete quarter packages and must not backfill actuals.
            return _local_non_gaap_actuals_allowed_for_source(src_name, path_in.name, txt)

        def _allow_non_gaap_metrics_from_local_page(src_name: str, path_in: Path, txt: str) -> bool:
            if not _allow_actuals_from_local_page(src_name, path_in, txt):
                return False
            if tkr_u == "ANF":
                # ANF financial schedules are parsed by the dedicated table parser
                # above. The generic relaxed OCR parser mis-scales several EX-99
                # HTML releases and creates noisy 1,000x adjusted metrics.
                return False
            return True

        def _detect_scale_txt(t: str) -> float:
            return _detect_local_non_gaap_text_scale(t)

        def _years_3m_from_text(lines: List[str]) -> List[int]:
            return _local_non_gaap_years_from_3m_lines(lines)

        def _slice_three_month_block_local(lines: List[str]) -> List[str]:
            return _local_non_gaap_three_month_lines(lines)

        def _three_month_end_from_text(txt: str) -> Optional[dt.date]:
            if not txt:
                return None
            # Prefer explicit "Three Months Ended <Month> <day>, <year>"
            m = re.search(r"three\s+months\s+ended\s+([A-Za-z]+)\s+(\d{1,2}),?\s*(\d{4})", txt, re.I)
            if m:
                try:
                    return _resolve_local_period_end(pd.Timestamp(f"{m.group(1)} {m.group(2)} {m.group(3)}").date())
                except Exception:
                    pass
            m_week = re.search(
                r"(?:thirteen|twenty[-\s]?six|thirty[-\s]?nine|fifty[-\s]?two|fifty[-\s]?three)\s+weeks\s+ended\s+([A-Za-z]+)\s+(\d{1,2}),?\s*(\d{4})",
                txt,
                re.I,
            )
            if m_week:
                try:
                    return _resolve_local_period_end(pd.Timestamp(f"{m_week.group(1)} {m_week.group(2)} {m_week.group(3)}").date())
                except Exception:
                    pass
            # If year is not on the same line, try to infer from nearby year headers
            m2 = re.search(r"three\s+months\s+ended\s+([A-Za-z]+)\s+(\d{1,2})", txt, re.I)
            if m2:
                years = [int(y) for y in re.findall(r"(20\d{2})", txt[:800])]
                if years:
                    try:
                        y = max(years)
                        return _resolve_local_period_end(pd.Timestamp(f"{m2.group(1)} {m2.group(2)} {y}").date())
                    except Exception:
                        pass
            return None

        def _pick_num_by_year(nums: List[float], years: List[int], q_end: Optional[dt.date]) -> Optional[float]:
            if not nums:
                return None
            if q_end is None or not years or len(nums) < 2:
                return nums[0]
            y = int(q_end.year)
            if y == years[0]:
                return nums[0]
            if len(years) > 1 and y == years[1]:
                return nums[1]
            return nums[0]

        def _parse_fcf_from_text(txt: str, q_end: Optional[dt.date]) -> Optional[float]:
            if not txt:
                return None
            lines = [re.sub(r"\s+", " ", ln).strip() for ln in txt.splitlines() if ln.strip()]
            lines_3m = _slice_three_month_block_local(lines)
            years = _years_3m_from_text(lines_3m)
            scale = _detect_scale_txt(txt)
            def _nums(line: str) -> List[float]:
                tokens = re.findall(r"\(?-?\d{1,3}(?:,\d{3})*(?:\.\d+)?\)?", line)
                nums: List[float] = []
                for t in tokens:
                    v = coerce_number(t)
                    if v is None:
                        continue
                    if 1900 <= float(v) <= 2100 and len(str(int(v))) == 4:
                        continue
                    nums.append(float(v) * scale)
                return nums

            for i, ln in enumerate(lines_3m):
                if "free cash flow" not in ln.lower():
                    continue
                nums = _nums(ln)
                if not nums:
                    for j in range(1, 3):
                        if i + j < len(lines_3m):
                            nums = _nums(lines_3m[i + j])
                            if nums:
                                break
                return _pick_num_by_year(nums, years, q_end)
            return None

        def _expand_quarter_ends(txt: str, q_end: Optional[dt.date]) -> List[dt.date]:
            if q_end is None:
                return []
            try:
                lines = [re.sub(r"\s+", " ", ln).strip() for ln in (txt or "").splitlines() if ln.strip()]
                lines_3m = _slice_three_month_block_local(lines)
                years = _years_3m_from_text(lines_3m)
            except Exception:
                years = []
            outs: List[dt.date] = []
            if years:
                alias_month_day: Optional[Tuple[int, int]] = None
                if q_end.month in (1, 2):
                    alias_month_day = (12, 31)
                elif q_end.month in (4, 5):
                    alias_month_day = (3, 31)
                elif q_end.month in (7, 8):
                    alias_month_day = (6, 30)
                elif q_end.month in (10, 11):
                    alias_month_day = (9, 30)
                for y in years:
                    try:
                        if local_period_aliases and alias_month_day:
                            aliased = dt.date(int(y), int(alias_month_day[0]), int(alias_month_day[1]))
                            outs.append(local_period_aliases.get(aliased, aliased))
                        else:
                            outs.append(dt.date(int(y), q_end.month, q_end.day))
                    except Exception:
                        continue
            if not outs:
                outs = [q_end]
            # preserve order (current-year first), unique
            seen: set[dt.date] = set()
            uniq: List[dt.date] = []
            for d in outs:
                if d not in seen:
                    uniq.append(d)
                    seen.add(d)
            return uniq

        def _missing_non_gaap_quarters(txt: str, q_end: Optional[dt.date]) -> List[dt.date]:
            return _filter_missing_local_non_gaap_metric_quarters(
                _expand_quarter_ends(txt, q_end),
                existing_metrics_by_quarter_for_local_fallback,
            )

        def _local_non_gaap_metric_quarters(txt: str, q_end: Optional[dt.date]) -> List[dt.date]:
            if tkr_u == "ANF":
                return _filter_missing_local_non_gaap_metric_quarters(
                    [q_end] if q_end is not None else [],
                    existing_metrics_by_quarter_for_local_fallback,
                )
            return _missing_non_gaap_quarters(txt, q_end)

        def _parse_segment_from_text(txt: str, q_end: Optional[dt.date]) -> List[Dict[str, Any]]:
            return _parse_local_non_gaap_segment_rows_from_text(txt, q_end)

        def _parse_debt_profile_from_text(txt: str, q_end: Optional[dt.date]) -> List[Dict[str, Any]]:
            if not txt:
                return []
            lines = [re.sub(r"\s+", " ", ln).strip() for ln in txt.splitlines() if ln.strip()]
            out: List[Dict[str, Any]] = []
            if not lines:
                return out
            scale = _detect_scale_txt(txt)

            def _amount_tokens(line_txt: str) -> List[str]:
                # Keep only amount-like columns ($000-style with commas) and dash placeholders.
                return re.findall(r"(?:\(?\d{1,3}(?:,\d{3})+\)?|[—–-])", line_txt)

            as_of_idx = 0
            as_of_match_found = False
            for ln in lines[:35]:
                dates = _parse_local_non_gaap_header_dates(ln)
                if len(dates) < 2:
                    continue
                if q_end is not None:
                    exact = [i for i, d in enumerate(dates) if abs((d - q_end).days) <= 7]
                    if exact:
                        as_of_idx = int(exact[0])
                        as_of_match_found = True
                        break
                    same_year = [i for i, d in enumerate(dates) if d.year == q_end.year]
                    if same_year:
                        as_of_idx = int(same_year[0])
                        as_of_match_found = True
                        break
                as_of_idx = 0
                as_of_match_found = True
                break

            for ln in lines:
                l = ln.lower()
                amt_cols = _amount_tokens(ln)
                if "principal amount" in l and amt_cols:
                    if as_of_idx < len(amt_cols):
                        v = coerce_number(amt_cols[as_of_idx])
                        if v is not None and float(v) > 0:
                            out.append(
                                {
                                    "quarter": q_end,
                                    "tranche": "Principal amount",
                                    "amount": float(v) * scale,
                                    "maturity_year": None,
                                    "unit": "USD",
                                    "is_table_total": True,
                                    "asof_col_idx": as_of_idx,
                                    "asof_match_found": as_of_match_found,
                                }
                            )
                    continue
                if len(ln) > 250:
                    continue
                if not re.search(r"\bdue\b", l):
                    continue
                if not re.search(r"\b(20\d{2})\b", l):
                    continue
                if not amt_cols:
                    continue
                if as_of_idx >= len(amt_cols):
                    continue
                tok = amt_cols[as_of_idx]
                if tok in {"-", "—", "–"}:
                    continue
                v = coerce_number(tok)
                if v is None or float(v) <= 0:
                    continue
                # Ensure this is a real debt instrument row, not textual boilerplate.
                if not re.search(r"(term\s+loan|notes?\s+due|convertible|debentures?)", l):
                    continue
                m = re.search(r"\b(20\d{2})\b", ln)
                my = int(m.group(1)) if m else None
                out.append(
                    {
                        "quarter": q_end,
                        "tranche": ln[:180],
                        "amount": float(v) * scale,
                        "maturity_year": my,
                        "unit": "USD",
                        "is_table_total": False,
                        "asof_col_idx": as_of_idx,
                        "asof_match_found": as_of_match_found,
                    }
                )
            return out

        def _parse_guidance_from_text(txt: str, q_end: Optional[dt.date]) -> List[Dict[str, Any]]:
            if not txt:
                return []
            lines = [re.sub(r"\s+", " ", ln).strip() for ln in txt.splitlines() if ln.strip()]
            out: List[Dict[str, Any]] = []
            if not lines:
                return out
            if tkr_u == "ANF":
                return _parse_anf_guidance_rows_from_lines(lines, q_end)
            anchor_re = re.compile(
                r"\b(guidance|outlook|financial outlook|updated guidance|full[- ]year outlook|next[- ]year outlook|targets?)\b",
                re.I,
            )
            intent_re = re.compile(
                r"\b(guidance|outlook|expect|expects|forecast|target|targets|plan|plans|intend|anticipate|reaffirm|maintain|raise|lower)\b",
                re.I,
            )
            metric_terms: Dict[str, List[str]] = {
                "Revenue": ["revenue", "sales", "top line"],
                "Adj EBITDA": ["adjusted ebitda", "adj ebitda", "ebitda"],
                "Adj EPS": [
                    "adjusted eps",
                    "adj eps",
                    "adjusted diluted eps",
                    "adjusted diluted earnings per share",
                    "adjusted earnings per share",
                    "earnings per share",
                    "eps",
                ],
                "FCF": ["free cash flow", "fcf", "free cash flow excluding"],
                "Capex": ["capex", "capital expenditures", "capital spending"],
                "Cost savings": ["cost savings", "savings", "run-rate savings", "annualized savings"],
                "Restructuring charges": ["restructuring", "transformation charges", "special items"],
                "Net debt / leverage": ["net leverage", "leverage", "net debt", "debt/ebitda"],
            }
            anti_re = re.compile(
                r"\b(performance obligations|transaction price allocated|recognized as follows|"
                r"securities act|registration exempt|forward-looking statements|safe harbor|"
                r"private securities litigation reform act|pslra|costs are amortized in a manner consistent|"
                r"contract performance period|not anticipated to be material|not expected to be material|"
                r"expected credit losses|reasonable and supportable forecast)\b",
                re.I,
            )
            num_re = re.compile(
                r"(\$?\s*[0-9]{1,4}(?:,[0-9]{3})*(?:\.[0-9]+)?\s*(?:bn|billion|m|million|%|bps|x)?"
                r"(?:\s*(?:to|through|\-|–|—)\s*\$?\s*[0-9]{1,4}(?:,[0-9]{3})*(?:\.[0-9]+)?\s*(?:bn|billion|m|million|%|bps|x)?)?)",
                re.I,
            )

            def _metric_hint(line_low: str) -> str:
                if "revenue" in line_low or "sales" in line_low or "top line" in line_low:
                    return "Revenue"
                if "cost savings" in line_low:
                    return "Cost savings"
                if "free cash flow" in line_low or re.search(r"\bfcf\b", line_low):
                    return "FCF"
                if "capex" in line_low or "capital expenditures" in line_low:
                    return "Capex"
                for m_name, kws in metric_terms.items():
                    if any(kw in line_low for kw in kws):
                        return m_name
                return "Other"

            def _num_tokens(line_txt: str) -> List[Tuple[str, float]]:
                out_nums: List[Tuple[str, float]] = []
                for m in num_re.finditer(line_txt):
                    raw = str(m.group(0) or "").strip()
                    raw_clean = raw.replace("$", "").replace(",", "").replace("%", "").replace("x", "").replace("bps", "")
                    try:
                        v = float(raw_clean)
                    except Exception:
                        continue
                    # Drop standalone years used as headers (2025/2026 etc.).
                    if re.fullmatch(r"\d{4}", raw_clean) and 1900 <= int(float(raw_clean)) <= 2100:
                        continue
                    out_nums.append((raw, v))
                return out_nums

            def _extract_low_high_metric_rows(line_txt: str, near_anchor_idx: Optional[int]) -> List[Tuple[str, str, str, Optional[str]]]:
                out_rows: List[Tuple[str, str, str, Optional[str]]] = []
                if not line_txt:
                    return out_rows
                ll = line_txt.lower()
                if not re.search(r"\blow\b.{0,40}\bhigh\b", ll):
                    return out_rows
                work = line_txt
                m_outlook = re.search(r"(full[- ]year\s+outlook|guidance|outlook)", work, re.I)
                if m_outlook:
                    work = work[m_outlook.start():]
                m_lh = re.search(r"\blow\b.{0,20}\bhigh\b", work, re.I)
                if m_lh:
                    work = work[m_lh.end():]
                year_hint = None
                for probe in [line_txt, lines[near_anchor_idx] if near_anchor_idx is not None and 0 <= near_anchor_idx < len(lines) else ""]:
                    m_y = re.search(r"\b(20\d{2})\b", str(probe or ""))
                    if m_y:
                        year_hint = m_y.group(1)
                        break

                metric_map = [
                    ("Revenue", r"revenue|sales|top line"),
                    ("Adj EBIT", r"adjusted\s+ebit|adj\.?\s+ebit"),
                    ("Adj EBITDA", r"adjusted\s+ebitda|adj\.?\s+ebitda"),
                    ("Adj EPS", r"adjusted\s+eps|adj\.?\s+eps|earnings\s+per\s+share|eps"),
                    ("FCF", r"free\s+cash\s+flow|\bfcf\b"),
                    ("Capex", r"capex|capital expenditures|capital spending"),
                    ("Cost savings", r"cost savings|savings"),
                ]
                for metric_name, mpat in metric_map:
                    pat = re.compile(
                        rf"(?:{mpat})[\s:\-,$()%]{{0,16}}\$?\s*([0-9]{{1,4}}(?:,[0-9]{{3}})*(?:\.[0-9]+)?)\s*(bn|billion|m|million|%|bps|x)?"
                        rf"[\s:\-,$()%]{{0,12}}\$?\s*([0-9]{{1,4}}(?:,[0-9]{{3}})*(?:\.[0-9]+)?)\s*(bn|billion|m|million|%|bps|x)?",
                        re.I,
                    )
                    mm = pat.search(work)
                    if not mm:
                        continue
                    lo = str(mm.group(1) or "")
                    hi = str(mm.group(3) or "")
                    unit = str(mm.group(2) or mm.group(4) or "").lower()
                    lo_disp = f"${lo}" if unit not in {"%", "bps", "x"} else f"{lo}{unit}"
                    hi_disp = f"${hi}" if unit not in {"%", "bps", "x"} else f"{hi}{unit}"
                    per = f" for FY {year_hint}" if year_hint else ""
                    out_rows.append((f"{metric_name} guidance {lo_disp} to {hi_disp}{per}", lo_disp, hi_disp, year_hint))
                return out_rows

            def _extract_metric_range_row(line_txt: str, near_anchor_idx: Optional[int]) -> List[Tuple[str, str, str, Optional[str]]]:
                line_clean = re.sub(r"[*†‡]", "", str(line_txt or ""))
                ll = line_clean.lower()
                metric_map = [
                    ("Revenue", r"revenue|sales|top line"),
                    ("Adj EBIT", r"adjusted\s+ebit|adj\.?\s+ebit"),
                    ("Adj EBITDA", r"adjusted\s+ebitda|adj\.?\s+ebitda"),
                    ("Adj EPS", r"adjusted\s+eps|adj\.?\s+eps|earnings\s+per\s+share|eps"),
                    ("FCF", r"free\s+cash\s+flow|\bfcf\b"),
                    ("Capex", r"capex|capital expenditures|capital spending"),
                    ("Cost savings", r"cost savings|savings"),
                ]
                metric_name = ""
                for cand_name, mpat in metric_map:
                    if re.search(rf"\b(?:{mpat})\b", ll, re.I):
                        metric_name = cand_name
                        break
                if not metric_name:
                    return []
                range_pat = re.compile(
                    r"(\$?\s*[0-9]{1,4}(?:,[0-9]{3})*(?:\.[0-9]+)?\s*(?:bn|billion|m|million|%|bps|x)?)"
                    r"\s*(?:to|through|\-|–|—)\s*"
                    r"(\$?\s*[0-9]{1,4}(?:,[0-9]{3})*(?:\.[0-9]+)?\s*(?:bn|billion|m|million|%|bps|x)?)",
                    re.I,
                )
                matches = list(range_pat.finditer(line_clean))
                if not matches:
                    return []
                first = matches[0]
                lo_disp = re.sub(r"\s+", "", str(first.group(1) or ""))
                hi_disp = re.sub(r"\s+", "", str(first.group(2) or ""))
                if lo_disp.startswith("$") and not hi_disp.startswith("$") and not re.search(r"%|bps|x$", hi_disp, re.I):
                    hi_disp = f"${hi_disp}"
                year_hint = None
                for probe in [line_clean, lines[near_anchor_idx] if near_anchor_idx is not None and 0 <= near_anchor_idx < len(lines) else ""]:
                    m_y = re.search(r"\b(20\d{2})\b", str(probe or ""))
                    if m_y:
                        year_hint = m_y.group(1)
                        break
                per = f" for FY {year_hint}" if year_hint else ""
                return [(f"{metric_name} guidance {lo_disp} to {hi_disp}{per}", lo_disp, hi_disp, year_hint)]

            anchor_idx: List[int] = [i for i, ln in enumerate(lines) if anchor_re.search(ln)]
            seen: set[str] = set()
            for i, ln in enumerate(lines):
                ll = ln.lower()
                if anti_re.search(ll):
                    continue
                if re.search(r"\$\s*change|%\s*change", ll):
                    continue
                near_anchor = any(abs(i - j) <= 18 for j in anchor_idx)
                has_intent = bool(intent_re.search(ll))
                metric_hint = _metric_hint(ll)
                num_parts = _num_tokens(ln)
                nums = [x[0] for x in num_parts]
                has_numeric = len(nums) > 0
                nearest_anchor = min(anchor_idx, key=lambda j: abs(i - j)) if (near_anchor and anchor_idx) else None

                # Handle "Low / High" outlook tables where ranges are listed without "to".
                table_rows = _extract_low_high_metric_rows(ln, nearest_anchor)
                if not table_rows:
                    table_rows = _extract_metric_range_row(ln, nearest_anchor)
                if table_rows and (near_anchor or has_intent):
                    for row_text, lo_disp, hi_disp, _yh in table_rows:
                        line_disp = row_text[:320]
                        key = re.sub(r"\s+", " ", line_disp).strip().lower()
                        if key in seen:
                            continue
                        seen.add(key)
                        out.append(
                            {
                                "quarter": q_end,
                                "line": line_disp,
                                "numbers": f"{lo_disp}, {hi_disp}",
                                "metric_hint": _metric_hint(line_disp.lower()),
                            }
                        )
                    continue

                if not has_numeric and i + 1 < len(lines):
                    nxt = lines[i + 1]
                    nxt_low = nxt.lower()
                    if not anti_re.search(nxt_low):
                        nxt_parts = _num_tokens(nxt)
                        nxt_nums = [x[0] for x in nxt_parts]
                        if nxt_nums and (metric_hint != "Other" or has_intent or near_anchor):
                            ln = f"{ln} {nxt}"
                            ll = ln.lower()
                            nums = nxt_nums
                            num_parts = nxt_parts
                            has_numeric = True

                if not has_numeric:
                    continue
                if not (has_intent or near_anchor):
                    continue
                if metric_hint == "Other" and not (has_intent or near_anchor):
                    continue
                if re.search(r"\bprovides?\s+the\s+following\s+guidance\s+for\b", ll) and metric_hint == "Other":
                    continue
                if len(re.findall(r"[A-Za-z]", ln)) < 16:
                    continue

                # Guidance tables often show "Metric | Low | High" without "to/-".
                if metric_hint != "Other" and len(nums) >= 2 and not re.search(r"\b(to|through|between)\b|\-|\u2013|\u2014", ln):
                    year_hint = None
                    nearest_anchor = None
                    if near_anchor and anchor_idx:
                        nearest_anchor = min(anchor_idx, key=lambda j: abs(i - j))
                    if nearest_anchor is not None:
                        m_y = re.search(r"\b(20\d{2})\b", lines[nearest_anchor])
                        if m_y:
                            year_hint = m_y.group(1)
                    n1 = float(num_parts[0][1]) if num_parts else None
                    n2 = float(num_parts[1][1]) if len(num_parts) > 1 else None
                    if (
                        metric_hint in {"Revenue", "Adj EBITDA", "FCF", "Capex", "Cost savings", "Restructuring charges"}
                        and n1 is not None
                        and n2 is not None
                        and abs(n1) < 10
                        and abs(n2) > 50
                    ):
                        continue
                    if year_hint and not re.search(r"\b20\d{2}\b", ln):
                        ln = f"{metric_hint} guidance {nums[0]} to {nums[1]} for FY {year_hint}"
                    else:
                        ln = f"{metric_hint} guidance {nums[0]} to {nums[1]}"

                line_disp = ln[:320]
                key = re.sub(r"\s+", " ", line_disp).strip().lower()
                if key in seen:
                    continue
                seen.add(key)
                out.append(
                    {
                        "quarter": q_end,
                        "line": line_disp,
                        "numbers": ", ".join(nums[:8]),
                        "metric_hint": metric_hint,
                    }
                )
            return out

        def _page_is_recon(txt: str) -> bool:
            t = (txt or "").lower()
            if "adjusted segment" in t or "reportable segments" in t:
                return False
            if "schedule of non-gaap financial measures" in t and (
                "adjusted non-gaap" in t or "adjusted ebitda" in t
            ):
                return True
            if "reconciliation of reported net income" in t and "adjusted ebitda" in t:
                return True
            if "reconciliation of reported consolidated results" in t and "adjusted ebitda" in t:
                return True
            if "reconciliation of reported" in t and "adjusted ebitda" in t:
                return True
            if "adjusted ebitda" in t and "adjusted ebit" in t and "reported net income" in t:
                return True
            if "adjusted ebitda" in t and "adjusted net income" in t:
                return True
            # Fallback: adjusted EBITDA + net income on the same page (common in slides)
            if "adjusted ebitda" in t and ("net income" in t or "net loss" in t) and ("adjusted ebit" in t or "adjusted net income" in t):
                return True
            return False

        def _ocr_page_text(page, cache_key: Optional[str] = None, cache_dir: Optional[Path] = None) -> str:
            try:
                import pytesseract  # type: ignore
                from PIL import Image  # type: ignore
            except Exception:
                return ""
            cache_path = None
            if cache_key:
                try:
                    cache_dir_use = cache_dir or (preferred_ticker_cache_root_from_base_dir(base_dir) / "slides_ocr")
                    cache_dir_use.mkdir(parents=True, exist_ok=True)
                    cache_path = cache_dir_use / f"{cache_key}.txt"
                    if cache_path.exists():
                        return cache_path.read_text(encoding="utf-8", errors="ignore")
                except Exception:
                    cache_path = None
            try:
                im = page.to_image(resolution=300).original
            except Exception:
                return ""
            try:
                txt = pytesseract.image_to_string(im)
                if cache_path is not None and txt:
                    try:
                        cache_path.write_text(txt, encoding="utf-8", errors="ignore")
                    except Exception:
                        pass
                return txt
            except Exception:
                return ""

        pages_per_q: Dict[pd.Timestamp, int] = {}
        rows_m_candidates: List[Dict[str, Any]] = []
        pdf_manifest_path = preferred_ticker_cache_root_from_base_dir(base_dir) / "local_non_gaap_pdf_manifest.json"
        pdf_manifest: Dict[str, Any] = {"version": 1, "files": {}}
        try:
            if pdf_manifest_path.exists():
                pdf_manifest = json.loads(pdf_manifest_path.read_text(encoding="utf-8", errors="ignore"))
        except Exception:
            pdf_manifest = {"version": 1, "files": {}}
        for src_name, folder in sources:
            if not folder.exists():
                continue
            files = sorted(folder.rglob("*"))
            for p in files[:200]:
                if not p.is_file():
                    continue
                if tkr_u == "ANF" and src_name == "slides" and p.suffix.lower() == ".xlsx":
                    try:
                        hist_rows = _parse_anf_quarterly_history_retail_driver_rows(
                            p,
                            fiscal_periods=anf_fiscal_map,
                        )
                    except Exception:
                        hist_rows = pd.DataFrame()
                    if hist_rows is not None and not hist_rows.empty:
                        rows_seg.extend(hist_rows.to_dict("records"))
                    continue
                if p.suffix.lower() not in (".txt", ".htm", ".html", ".pdf"):
                    continue
                # For slides, skip very old decks outside the configured window (perf).
                if src_name == "slides" and config.min_year:
                    q_hint = _infer_q_from_filename(p.name)
                    if q_hint is not None and q_hint.year < int(config.min_year):
                        continue
                if tkr_u == "ANF" and src_name in {"slides", "earnings_release"}:
                    name_low = p.name.lower()
                    if "financial_schedules" in name_low or name_low.startswith("8-k_") or "earnings_release" in name_low:
                        try:
                            full_text, full_lines = _anf_extract_material_lines(
                                p,
                                cache_root=config.cache_dir,
                                rebuild_cache=config.rebuild_doc_text_cache,
                                quiet_pdf_warnings=config.quiet_pdf_warnings,
                            )
                        except Exception:
                            full_text, full_lines = "", []
                        q_full = _infer_anf_quarter_from_material(
                            p,
                            full_text,
                            fiscal_map=anf_fiscal_map,
                            aliases=local_period_aliases,
                        )
                        if q_full is not None and full_lines:
                            scale_full = _detect_local_non_gaap_text_scale(full_text)
                            anf_rows = _parse_anf_adjusted_metrics_from_lines(
                                full_lines,
                                quarter_end=q_full,
                                scale=scale_full,
                                source_doc=str(p),
                                source=src_name,
                            )
                            if anf_rows:
                                rows_m_candidates.extend(anf_rows)
                                rows_f.append(
                                    {
                                        "accn": None,
                                        "filed": None,
                                        "status": "ok_anf_financial_schedule",
                                        "doc": str(p),
                                        "quarter": str(q_full),
                                        "col": "ANF financial schedule",
                                        "source": src_name,
                                        "page": None,
                                    }
                                )
                            anf_guidance_rows = _parse_anf_guidance_rows_from_lines(full_lines, q_full)
                            if anf_guidance_rows:
                                for r0 in anf_guidance_rows:
                                    r0.update({"doc": str(p), "page": None, "source": src_name})
                                rows_guid.extend(anf_guidance_rows)
                            if p.suffix.lower() in {".htm", ".html"}:
                                mix_rows = _parse_anf_sales_mix_tables_from_html(p, q_full)
                                if mix_rows is not None and not mix_rows.empty:
                                    rows_seg.extend(mix_rows.to_dict("records"))
                            retail_text_rows = _parse_anf_retail_text_driver_rows_from_lines(
                                full_lines,
                                quarter_end=q_full,
                                source_doc=str(p),
                                source_type=src_name,
                            )
                            if retail_text_rows is not None and not retail_text_rows.empty:
                                rows_seg.extend(retail_text_rows.to_dict("records"))
                if p.suffix.lower() == ".pdf":
                    try:
                        import pdfplumber  # type: ignore
                    except Exception:
                        continue
                    try:
                        use_cache_only = False
                        cached_pages = None
                        text_cache_dir = None
                        ocr_cache_dir = None
                        try:
                            st = p.stat()
                            try:
                                manifest_key = str(p.resolve())
                            except Exception:
                                manifest_key = str(p)
                            entry = (pdf_manifest.get("files", {}) or {}).get(manifest_key)
                            if entry and entry.get("mtime") == st.st_mtime and entry.get("size") == st.st_size and entry.get("pages"):
                                cached_pages = int(entry.get("pages"))
                                use_cache_only = True
                        except Exception:
                            use_cache_only = False
                        try:
                            text_cache_dir, ocr_cache_dir = _local_non_gaap_pdf_cache_dirs(base_dir, src_name)
                            text_cache_dir.mkdir(parents=True, exist_ok=True)
                            ocr_cache_dir.mkdir(parents=True, exist_ok=True)
                        except Exception:
                            text_cache_dir = None
                            ocr_cache_dir = None

                        def _read_cached_page(cache_key: str) -> str:
                            for d in (text_cache_dir, ocr_cache_dir):
                                if d is None:
                                    continue
                                try:
                                    pth = d / f"{cache_key}.txt"
                                    if pth.exists():
                                        return pth.read_text(encoding="utf-8", errors="ignore")
                                except Exception:
                                    continue
                            return ""

                        if use_cache_only and cached_pages:
                            for idx in range(cached_pages):
                                cache_key = _local_non_gaap_pdf_cache_key(p, src_name=src_name, page_number=idx + 1)
                                txt = _read_cached_page(cache_key)
                                if not txt:
                                    continue
                                scores = _local_non_gaap_page_scores(txt)
                                if max(scores.values()) == 0:
                                    continue
                                q_end = _resolve_local_period_end(
                                    _three_month_end_from_text(txt) or _infer_q_from_filename(p.name) or infer_quarter_end_from_text(txt)
                                )
                                if q_end is None:
                                    continue
                                actuals_allowed = _allow_actuals_from_local_page(src_name, p, txt)
                                if _allow_non_gaap_metrics_from_local_page(src_name, p, txt) and scores.get("non_gaap", 0) >= 2 and _page_is_recon(txt):
                                    for q_end_use in _local_non_gaap_metric_quarters(txt, q_end):
                                        q_ts = pd.Timestamp(q_end_use)
                                        if pages_per_q.get(q_ts, 0) >= 2:
                                            continue
                                        aebit, aebitda, aeps, adj, status, col_label = parse_adjusted_from_plain_text(txt, q_end_use, mode="relaxed")
                                        if status in ("ok_ocr", "ok_relaxed_ocr"):
                                            fcf_val = _parse_fcf_from_text(txt, q_end_use)
                                            rows_m_candidates.append({
                                                "quarter": q_end_use,
                                                "adj_ebit": aebit,
                                                "adj_ebitda": aebitda,
                                                "adj_eps": aeps,
                                                "adj_fcf": fcf_val,
                                                "source": src_name,
                                                "source_type": "earnings_deck",
                                                "accn": None,
                                                "filed": None,
                                                "doc": str(p),
                                                "page": idx + 1,
                                                "confidence": "low",
                                                "col": col_label,
                                                "source_snippet": "Adjusted EBITDA row",
                                                "score": scores.get("non_gaap", 0),
                                            })
                                            rows_f.append({
                                                "accn": None,
                                                "filed": None,
                                                "status": "ok_local",
                                                "doc": str(p),
                                                "quarter": str(q_end_use),
                                                "col": col_label,
                                                "source": src_name,
                                                "page": idx + 1,
                                            })
                                            seen_q.add(q_ts)
                                            pages_per_q[q_ts] = pages_per_q.get(q_ts, 0) + 1
                                if actuals_allowed and scores.get("segment", 0) >= 2:
                                    seg_rows = _parse_segment_from_text(txt, q_end)
                                    if seg_rows:
                                        for r0 in seg_rows:
                                            r0.update({"doc": str(p), "page": idx + 1, "source": src_name})
                                            if src_name == "annual_reports":
                                                r0.update({"period_type": "annual", "source_period_label": "annual"})
                                        rows_seg.extend(seg_rows)
                                if actuals_allowed and scores.get("debt", 0) >= 2 and _local_non_gaap_debt_source_allowed(
                                    src_name,
                                    has_financial_statement_files=has_financial_statement_files,
                                ):
                                    debt_rows = []
                                    if src_name == "financial_statement" and p.suffix.lower() in {".htm", ".html"}:
                                        debt_rows = _parse_financial_statement_debt_table_html(p, q_end)
                                    if not debt_rows:
                                        debt_rows = _parse_debt_profile_from_text(txt, q_end)
                                    if debt_rows:
                                        for r0 in debt_rows:
                                            r0.update({"doc": str(p), "page": idx + 1, "source": src_name})
                                        rows_debt.extend(debt_rows)
                                if scores.get("guidance", 0) >= 2:
                                    guid_rows = _parse_guidance_from_text(txt, q_end)
                                    if guid_rows:
                                        for r0 in guid_rows:
                                            r0.update({"doc": str(p), "page": idx + 1, "source": src_name})
                                        rows_guid.extend(guid_rows)
                            continue
                        with silence_pdfminer_warnings(enabled=config.quiet_pdf_warnings):
                            with pdfplumber.open(str(p)) as pdf:
                                n_pages = len(pdf.pages)
                                for idx, page in enumerate(pdf.pages):
                                    txt = ""
                                    cache_key = _local_non_gaap_pdf_cache_key(p, src_name=src_name, page_number=idx + 1)
                                    if text_cache_dir is not None:
                                        try:
                                            cache_path = text_cache_dir / f"{cache_key}.txt"
                                            if cache_path.exists():
                                                txt = cache_path.read_text(encoding="utf-8", errors="ignore")
                                        except Exception:
                                            txt = ""
                                    if not txt:
                                        txt = page.extract_text() or ""
                                        if text_cache_dir is not None:
                                            try:
                                                cache_path = text_cache_dir / f"{cache_key}.txt"
                                                if not cache_path.exists():
                                                    cache_path.write_text(txt or "", encoding="utf-8", errors="ignore")
                                            except Exception:
                                                pass
                                    # text-first, OCR only if low text
                                    if txt is None:
                                        txt = ""
                                    hint_txt = txt.lower()
                                    has_hint = any(
                                        k in hint_txt
                                        for k in (
                                            "adjusted ebitda",
                                            "adjusted ebit",
                                            "reconciliation",
                                            "adjusted diluted earnings",
                                            "free cash flow",
                                            "appendix: financial information",
                                        )
                                    )
                                    if len(txt.strip()) < 200:
                                        # OCR only if the page is likely relevant or near the end (slides appendices).
                                        if src_name == "slides" and (has_hint or idx >= max(0, n_pages - 6)):
                                            ocr_txt = _ocr_page_text(page, cache_key=cache_key, cache_dir=ocr_cache_dir)
                                            if ocr_txt and len(ocr_txt) > len(txt):
                                                txt = ocr_txt
                                        elif src_name != "slides" and has_hint:
                                            ocr_txt = _ocr_page_text(page, cache_key=cache_key, cache_dir=ocr_cache_dir)
                                            if ocr_txt and len(ocr_txt) > len(txt):
                                                txt = ocr_txt
                                    scores = _local_non_gaap_page_scores(txt)
                                    # For slides: OCR can recover key lines even when text exists but is low-signal.
                                    if src_name == "slides" and scores.get("non_gaap", 0) < 2 and (has_hint or idx >= max(0, n_pages - 6)):
                                        ocr_txt = _ocr_page_text(page, cache_key=cache_key, cache_dir=ocr_cache_dir)
                                        if ocr_txt and len(ocr_txt) > len(txt):
                                            txt = ocr_txt
                                            scores = _local_non_gaap_page_scores(txt)
                                    if max(scores.values()) == 0 and "appendix: financial information" in txt.lower():
                                        ocr_txt = _ocr_page_text(page, cache_key=cache_key, cache_dir=ocr_cache_dir)
                                        if ocr_txt and len(ocr_txt) > len(txt):
                                            txt = ocr_txt
                                            scores = _local_non_gaap_page_scores(txt)
                                    if max(scores.values()) == 0:
                                        continue
                                q_end = _three_month_end_from_text(txt) or _infer_q_from_filename(p.name) or infer_quarter_end_from_text(txt)
                                if q_end is None:
                                    continue
                                actuals_allowed = _allow_actuals_from_local_page(src_name, p, txt)
                                if _allow_non_gaap_metrics_from_local_page(src_name, p, txt) and scores.get("non_gaap", 0) >= 2 and _page_is_recon(txt):
                                    for q_end_use in _local_non_gaap_metric_quarters(txt, q_end):
                                        q_ts = pd.Timestamp(q_end_use)
                                        if pages_per_q.get(q_ts, 0) >= 2:
                                            continue
                                        aebit, aebitda, aeps, adj, status, col_label = parse_adjusted_from_plain_text(txt, q_end_use, mode="relaxed")
                                        if status in ("ok_ocr", "ok_relaxed_ocr"):
                                            fcf_val = _parse_fcf_from_text(txt, q_end_use)
                                            rows_m_candidates.append({
                                                "quarter": q_end_use,
                                                "adj_ebit": aebit,
                                                "adj_ebitda": aebitda,
                                                "adj_eps": aeps,
                                                "adj_fcf": fcf_val,
                                                "source": src_name,
                                                "source_type": "earnings_deck",
                                                "accn": None,
                                                "filed": None,
                                                "doc": str(p),
                                                "page": idx + 1,
                                                "confidence": "low",
                                                "col": col_label,
                                                "source_snippet": "Adjusted EBITDA row",
                                                "score": scores.get("non_gaap", 0),
                                            })
                                            rows_f.append({
                                                "accn": None,
                                                "filed": None,
                                                "status": "ok_local",
                                                "doc": str(p),
                                                "quarter": str(q_end_use),
                                                "col": col_label,
                                                "source": src_name,
                                                "page": idx + 1,
                                            })
                                            seen_q.add(q_ts)
                                            pages_per_q[q_ts] = pages_per_q.get(q_ts, 0) + 1
                                if actuals_allowed and scores.get("segment", 0) >= 2:
                                    seg_rows = _parse_segment_from_text(txt, q_end)
                                    if seg_rows:
                                        for r0 in seg_rows:
                                            r0.update({"doc": str(p), "page": idx + 1, "source": src_name})
                                            if src_name == "annual_reports":
                                                r0.update({"period_type": "annual", "source_period_label": "annual"})
                                        rows_seg.extend(seg_rows)
                                if actuals_allowed and scores.get("debt", 0) >= 2 and _local_non_gaap_debt_source_allowed(
                                    src_name,
                                    has_financial_statement_files=has_financial_statement_files,
                                ):
                                    debt_rows = []
                                    if src_name == "financial_statement" and p.suffix.lower() in {".htm", ".html"}:
                                        debt_rows = _parse_financial_statement_debt_table_html(p, q_end)
                                    if not debt_rows:
                                        debt_rows = _parse_debt_profile_from_text(txt, q_end)
                                    if debt_rows:
                                        for r0 in debt_rows:
                                            r0.update({"doc": str(p), "page": idx + 1, "source": src_name})
                                        rows_debt.extend(debt_rows)
                                if scores.get("guidance", 0) >= 2:
                                    guid_rows = _parse_guidance_from_text(txt, q_end)
                                    if guid_rows:
                                        for r0 in guid_rows:
                                            r0.update({"doc": str(p), "page": idx + 1, "source": src_name})
                                        rows_guid.extend(guid_rows)
                            try:
                                st = p.stat()
                                try:
                                    manifest_key = str(p.resolve())
                                except Exception:
                                    manifest_key = str(p)
                                pdf_manifest.setdefault("files", {})[manifest_key] = {
                                    "mtime": st.st_mtime,
                                    "size": st.st_size,
                                    "pages": n_pages,
                                }
                            except Exception:
                                pass
                    except Exception:
                        continue
                else:
                    txt = _extract_text_from_file(p)
                    if not txt or len(txt) < 50:
                        continue
                    scores = _local_non_gaap_page_scores(txt)
                    if max(scores.values()) == 0:
                        continue
                    q_end = _resolve_local_period_end(
                        _three_month_end_from_text(txt) or _infer_q_from_filename(p.name) or infer_quarter_end_from_text(txt)
                    )
                    if q_end is None:
                        continue
                    if tkr_u == "ANF":
                        retail_text_rows = _parse_anf_retail_text_driver_rows_from_lines(
                            [re.sub(r"\s+", " ", ln).strip() for ln in txt.splitlines() if ln.strip()],
                            quarter_end=q_end,
                            source_doc=str(p),
                            source_type=src_name,
                        )
                        if retail_text_rows is not None and not retail_text_rows.empty:
                            rows_seg.extend(retail_text_rows.to_dict("records"))
                    actuals_allowed = _allow_actuals_from_local_page(src_name, p, txt)
                    if _allow_non_gaap_metrics_from_local_page(src_name, p, txt) and scores.get("non_gaap", 0) >= 2 and _page_is_recon(txt):
                        for q_end_use in _local_non_gaap_metric_quarters(txt, q_end):
                            q_ts = pd.Timestamp(q_end_use)
                            if pages_per_q.get(q_ts, 0) >= 2:
                                continue
                            aebit, aebitda, aeps, adj, status, col_label = parse_adjusted_from_plain_text(txt, q_end_use, mode="relaxed")
                            if status in ("ok_ocr", "ok_relaxed_ocr"):
                                fcf_val = _parse_fcf_from_text(txt, q_end_use)
                                rows_m_candidates.append({
                                    "quarter": q_end_use,
                                    "adj_ebit": aebit,
                                    "adj_ebitda": aebitda,
                                    "adj_eps": aeps,
                                    "adj_fcf": fcf_val,
                                    "source": src_name,
                                    "source_type": "earnings_deck",
                                    "accn": None,
                                    "filed": None,
                                    "doc": str(p),
                                    "page": None,
                                    "confidence": "low",
                                    "col": col_label,
                                    "source_snippet": "Adjusted EBITDA row",
                                    "score": scores.get("non_gaap", 0),
                                })
                                rows_f.append({
                                    "accn": None,
                                    "filed": None,
                                    "status": "ok_local",
                                    "doc": str(p),
                                    "quarter": str(q_end_use),
                                    "col": col_label,
                                    "source": src_name,
                                    "page": None,
                                })
                                seen_q.add(q_ts)
                                pages_per_q[q_ts] = pages_per_q.get(q_ts, 0) + 1
                    if actuals_allowed and scores.get("segment", 0) >= 2:
                        seg_rows = _parse_segment_from_text(txt, q_end)
                        if seg_rows:
                            for r0 in seg_rows:
                                r0.update({"doc": str(p), "page": None, "source": src_name})
                                if src_name == "annual_reports":
                                    r0.update({"period_type": "annual", "source_period_label": "annual"})
                            rows_seg.extend(seg_rows)
                    if actuals_allowed and scores.get("debt", 0) >= 2 and _local_non_gaap_debt_source_allowed(
                        src_name,
                        has_financial_statement_files=has_financial_statement_files,
                    ):
                        debt_rows = []
                        if src_name == "financial_statement" and p.suffix.lower() in {".htm", ".html"}:
                            debt_rows = _parse_financial_statement_debt_table_html(p, q_end)
                        if not debt_rows:
                            debt_rows = _parse_debt_profile_from_text(txt, q_end)
                        if debt_rows:
                            for r0 in debt_rows:
                                r0.update({"doc": str(p), "page": None, "source": src_name})
                            rows_debt.extend(debt_rows)
                    if scores.get("guidance", 0) >= 2:
                        guid_rows = _parse_guidance_from_text(txt, q_end)
                        if guid_rows:
                            for r0 in guid_rows:
                                r0.update({"doc": str(p), "page": None, "source": src_name})
                            rows_guid.extend(guid_rows)
        # Deduplicate adjusted metrics per quarter without collapsing distinct metrics
        # into a single winning row. We keep the best available value per metric.
        df_m = pd.DataFrame(rows_m_candidates)
        if not df_m.empty and "quarter" in df_m.columns:
            df_m["quarter"] = pd.to_datetime(df_m["quarter"], errors="coerce")
            df_m = df_m[df_m["quarter"].notna()]
            if not df_m.empty:
                df_m["score"] = pd.to_numeric(df_m.get("score"), errors="coerce").fillna(0)
                if "period_type" not in df_m.columns:
                    df_m["period_type"] = "quarter"
                df_m["period_type"] = df_m["period_type"].astype(str).str.strip().str.lower().replace({"": "quarter"})
                period_order_map = {"annual": 0, "ytd": 1, "quarter": 2}
                df_m["_period_order"] = df_m["period_type"].map(period_order_map).fillna(2).astype(int)
                metric_cols = ["adj_ebit", "adj_ebitda", "adj_eps", "adj_fcf"]
                merged_rows: List[Dict[str, Any]] = []
                for (qv, period_type), sub in df_m.sort_values(["quarter", "_period_order", "score"], kind="stable").groupby(["quarter", "period_type"], sort=True):
                    sub = sub.copy()
                    for metric_col in metric_cols:
                        sub[f"{metric_col}_num"] = pd.to_numeric(sub.get(metric_col), errors="coerce")
                        sub[f"{metric_col}_nonnull"] = sub[f"{metric_col}_num"].notna().astype(int)
                        sub[f"{metric_col}_abs"] = sub[f"{metric_col}_num"].abs()
                    sub["_metric_count"] = sub[[f"{metric_col}_nonnull" for metric_col in metric_cols]].sum(axis=1)
                    base = (
                        sub.sort_values(["score", "_metric_count"], ascending=[False, False])
                        .iloc[0]
                        .to_dict()
                    )
                    merged = dict(base)
                    merged["quarter"] = qv
                    merged["period_type"] = period_type
                    merged["confidence"] = merged.get("confidence") or "low"
                    merged_sources: List[str] = []
                    for metric_col in metric_cols:
                        metric_sub = sub[sub[f"{metric_col}_nonnull"] > 0].copy()
                        if metric_sub.empty:
                            merged[metric_col] = pd.NA
                            continue
                        metric_best = (
                            metric_sub.sort_values(
                                ["score", f"{metric_col}_abs"],
                                ascending=[False, False],
                            )
                            .iloc[0]
                            .to_dict()
                        )
                        merged[metric_col] = metric_best.get(metric_col)
                        metric_doc = str(metric_best.get("doc") or "").strip()
                        metric_page = metric_best.get("page")
                        if metric_doc:
                            merged_sources.append(f"{metric_col}:{metric_doc}{f' p.{metric_page}' if pd.notna(metric_page) else ''}")
                    if merged_sources:
                        merged["source_snippet"] = "Merged local fallback metrics | " + " | ".join(merged_sources[:4])
                    merged_rows.append(merged)
                df_m = pd.DataFrame(merged_rows)
                if not df_m.empty and "period_type" in df_m.columns:
                    df_m["_period_order"] = df_m["period_type"].map(period_order_map).fillna(2).astype(int)
                    df_m = df_m.sort_values(["quarter", "_period_order", "score"], kind="stable").reset_index(drop=True)
                df_m = df_m.drop(
                    columns=[
                        *[f"{metric_col}_num" for metric_col in metric_cols],
                        *[f"{metric_col}_nonnull" for metric_col in metric_cols],
                        *[f"{metric_col}_abs" for metric_col in metric_cols],
                        "_metric_count",
                        "_period_order",
                    ],
                    errors="ignore",
                )
        try:
            if pdf_manifest_path:
                pdf_manifest_path.parent.mkdir(parents=True, exist_ok=True)
                pdf_manifest_path.write_text(json.dumps(pdf_manifest, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass
        df_debt = _limit_recent_financial_statement_debt_rows(pd.DataFrame(rows_debt))
        df_debt = _drop_financial_statement_debt_rows_covered_by_slides(df_debt)
        df_guid_raw = _dedupe_slides_guidance_rows(pd.DataFrame(rows_guid))
        df_guid = _normalize_anf_guidance_rows(df_guid_raw) if tkr_u == "ANF" else df_guid_raw
        return (
            df_m,
            pd.DataFrame(rows_f),
            _dedupe_local_non_gaap_segment_rows(pd.DataFrame(rows_seg)),
            df_debt,
            df_guid_raw,
            df_guid,
        )

    # Use local fallback only when EX-99 (strict/relaxed) missing for a quarter
    if config.enable_tier3_non_gaap:
        def _has_any_metric(df: pd.DataFrame) -> pd.Series:
            if df is None or df.empty:
                return pd.Series([], dtype=bool)
            cols = [c for c in LOCAL_NON_GAAP_CANONICAL_METRICS if c in df.columns]
            if not cols:
                return pd.Series([False] * len(df), index=df.index)
            return df[cols].notna().any(axis=1)
        # Only treat strict EX-99 metrics as canonical; relaxed preview should not block local fallback.
        canonical_adj_metrics = (
            adj_metrics.loc[_has_any_metric(adj_metrics)].copy()
            if adj_metrics is not None and not adj_metrics.empty
            else pd.DataFrame()
        )
        existing_q = _normalized_quarter_timestamps(
            canonical_adj_metrics["quarter"]
            if not canonical_adj_metrics.empty
            else pd.Series([], dtype="datetime64[ns]")
        )
        existing_metrics_by_quarter_for_local_fallback = _existing_local_non_gaap_metrics_by_quarter(
            canonical_adj_metrics
        )
        # The local fallback stage rescues adjusted metrics and slide-derived support
        # from curated local materials. It is keyed by local material signature so
        # new PDFs/TXTs invalidate the cache without touching SEC-driven stages.
        local_fallback_key = "|".join(
            [
                f"v{LOCAL_NON_GAAP_FALLBACK_VERSION}",
                f"materials={local_material_sig}",
                f"max_q={config.max_quarters}",
                f"quiet_pdf={int(bool(config.quiet_pdf_warnings))}",
                "doc_text_cache=v2",
            ]
        )
        local_fallback_cached = None if config.rebuild_doc_text_cache else _load_stage_cache("local_non_gaap_fallback", local_fallback_key)
        if isinstance(local_fallback_cached, dict):
            local_metrics = local_fallback_cached.get("metrics", pd.DataFrame())
            local_files = local_fallback_cached.get("files", pd.DataFrame())
            local_segments = local_fallback_cached.get("segments", pd.DataFrame())
            local_debt = local_fallback_cached.get("debt", pd.DataFrame())
            local_guidance_raw = local_fallback_cached.get("guidance_raw", pd.DataFrame())
            local_guidance = local_fallback_cached.get("guidance", pd.DataFrame())
        else:
            with _timed_stage(stage_timings, "local_non_gaap_fallback", enabled=config.profile_timings):
                local_metrics, local_files, local_segments, local_debt, local_guidance_raw, local_guidance = build_non_gaap_local_fallback()
            _save_stage_cache(
                "local_non_gaap_fallback",
                local_fallback_key,
                {
                    "metrics": local_metrics,
                    "files": local_files,
                    "segments": local_segments,
                    "debt": local_debt,
                    "guidance_raw": local_guidance_raw,
                    "guidance": local_guidance,
                },
            )
        if not local_metrics.empty:
            local_metrics = _prune_local_non_gaap_metrics_against_existing(
                local_metrics,
                existing_metrics_by_quarter_for_local_fallback,
            )
            if not local_metrics.empty:
                adj_metrics = pd.concat([adj_metrics, local_metrics], ignore_index=True)
                kept_local_quarters = _normalized_quarter_timestamps(local_metrics["quarter"])
                if kept_local_quarters and local_files is not None and not local_files.empty and "quarter" in local_files.columns:
                    local_files = local_files.copy()
                    local_files["quarter"] = pd.to_datetime(local_files["quarter"], errors="coerce")
                    local_files = local_files[local_files["quarter"].isin(kept_local_quarters)]
                if non_gaap_files is None or non_gaap_files.empty:
                    non_gaap_files = local_files
                else:
                    non_gaap_files = pd.concat([non_gaap_files, local_files], ignore_index=True)
        # stash local slide/segment/debt/guidance extracts for Excel (if any)
        slides_segments = local_segments
        slides_debt = local_debt
        slides_guidance_raw = local_guidance_raw
        slides_guidance = local_guidance
    if tkr_u == "ANF":
        adj_metrics, adj_breakdown = _sanitize_anf_adjusted_metric_units(adj_metrics, adj_breakdown, hist)

    non_gaap_needs = pd.DataFrame()
    non_gaap_info = pd.DataFrame()
    if config.non_gaap_mode == "strict":
        if non_gaap_files is not None and not non_gaap_files.empty:
            # Only flag truly ambiguous cases in strict mode (e.g., parse errors).
            needs_statuses = {"parse_error"}
            ng_warn = non_gaap_files[non_gaap_files["status"].astype(str).isin(needs_statuses)].copy()
            if not ng_warn.empty:
                # Non-GAAP file logs may not have a 'quarter' column; fall back to quarter_end if present.
                if "quarter" not in ng_warn.columns:
                    if "quarter_end" in ng_warn.columns:
                        ng_warn["quarter"] = ng_warn["quarter_end"]
                    else:
                        ng_warn["quarter"] = pd.NaT
                non_gaap_needs = ng_warn.assign(
                    metric="non_gaap",
                    severity="warn",
                    message="Non-GAAP parsing error.",
                    source=ng_warn["status"],
                )[["quarter", "metric", "severity", "message", "source"]]
            # In strict mode, "no_matching_column" is informational (strict empty is OK).
            ng_info = non_gaap_files[non_gaap_files["status"].astype(str).isin({"no_matching_column"})].copy()
            if not ng_info.empty:
                if "quarter" not in ng_info.columns:
                    if "quarter_end" in ng_info.columns:
                        ng_info["quarter"] = ng_info["quarter_end"]
                    else:
                        ng_info["quarter"] = pd.NaT
                non_gaap_info = ng_info.assign(
                    metric="non_gaap",
                    severity="info",
                    message="Non-GAAP strict: no matching column (expected for many files).",
                    source=ng_info["status"],
                )[["quarter", "metric", "severity", "message", "source"]]
    if config.non_gaap_mode == "relaxed" and not adj_metrics.empty:
        relaxed_warn = adj_metrics.assign(
            metric="non_gaap_relaxed",
            severity="warn",
            message="Non-GAAP relaxed mode (low confidence).",
            source="relaxed_mode",
        )[["quarter", "metric", "severity", "message", "source"]]
        non_gaap_needs = pd.concat([non_gaap_needs, relaxed_warn], ignore_index=True) if not non_gaap_needs.empty else relaxed_warn
    elif not adj_metrics_relaxed.empty:
        relaxed_warn = adj_metrics_relaxed.assign(
            metric="non_gaap_relaxed",
            severity="warn",
            message="Non-GAAP relaxed mode (low confidence).",
            source="relaxed_mode",
        )[["quarter", "metric", "severity", "message", "source"]]
        non_gaap_needs = pd.concat([non_gaap_needs, relaxed_warn], ignore_index=True) if not non_gaap_needs.empty else relaxed_warn

    issues1 = validate_history(hist)

    lt_debt_map = pd.DataFrame()
    if not debt_tranches.empty:
        lt_rows: List[Dict[str, Any]] = []
        for q in sorted(debt_tranches["quarter"].dropna().unique()):
            pr_lt = compute_long_term_debt_instant(df_all, end=q, prefer_forms=["10-Q", "10-K"])
            if pr_lt is not None and pr_lt.value is not None:
                lt_rows.append({"quarter": q, "long_term_debt": float(pr_lt.value)})
        if lt_rows:
            lt_debt_map = pd.DataFrame(lt_rows)

    issues2 = validate_debt_tieout(hist, debt_tranches, lt_debt_map)
    issues3 = needs_review_from_audit(audit)
    info_log = info_log_from_audit(audit)
    if non_gaap_info is not None and not non_gaap_info.empty:
        info_log = pd.concat([info_log, non_gaap_info], ignore_index=True) if not info_log.empty else non_gaap_info

    def _append_fy_q4_qa_rows(info_df: pd.DataFrame) -> pd.DataFrame:
        rows: List[Dict[str, Any]] = []
        if hist is None or hist.empty or "quarter" not in hist.columns:
            return info_df

        hq = hist.copy()
        hq["quarter"] = pd.to_datetime(hq["quarter"], errors="coerce")
        hq = hq[hq["quarter"].notna()].sort_values("quarter")
        if hq.empty:
            return info_df

        q_target = pd.Timestamp(dt.date(2025, 12, 31))
        if not (hq["quarter"] == q_target).any():
            rows.append(
                {
                    "quarter": q_target.date(),
                    "metric": "QA_FY_Q4",
                    "severity": "warn",
                    "message": "2025-12-31 row missing in History_Q.",
                    "source": "history",
                }
            )
            qa_df = pd.DataFrame(rows)
            return pd.concat([info_df, qa_df], ignore_index=True) if info_df is not None and not info_df.empty else qa_df

        q9 = pd.Timestamp(dt.date(2025, 9, 30))
        spec_map = {s.name: s for s in GAAP_SPECS}

        def _form_rank(v: Any) -> int:
            f = str(v or "").upper()
            if f.startswith("10-K"):
                return 0
            if f.startswith("10-Q"):
                return 1
            return 9

        def _pick_cf_duration(tags: List[str], end_d: dt.date, dur_class: str, preferred_tag: Optional[str] = None) -> Optional[float]:
            if not tags:
                return None
            cand = df_all[(df_all["tag"].isin(tags)) & (df_all["end_d"] == end_d) & df_all["start_d"].notna()].copy()
            if cand.empty:
                return None
            spec_tmp = MetricSpec("tmp", tags, "duration", "USD", ["10-K", "10-Q"])
            cand = _filter_unit(cand, spec_tmp)
            if cand.empty:
                return None
            cand["dur"] = _duration_days(cand["end_d"], cand["start_d"])
            cand["dur_class"] = cand["dur"].apply(classify_duration)
            cand = cand[cand["dur_class"] == dur_class].copy()
            if cand.empty:
                return None
            pref_tag = str(preferred_tag or "").strip()
            if pref_tag:
                pref = cand[cand["tag"].astype(str) == pref_tag].copy()
                if not pref.empty:
                    cand = pref
            cand["form_rank"] = cand["form"].apply(_form_rank)
            cand = cand.sort_values(["form_rank", "filed_d"], ascending=[True, False])
            try:
                return float(cand.iloc[0]["val"])
            except Exception:
                return None

        def _pick_cf_instant(tags: List[str], end_d: dt.date) -> Optional[float]:
            if not tags:
                return None
            cand = df_all[(df_all["tag"].isin(tags)) & (df_all["end_d"] == end_d)].copy()
            if cand.empty:
                return None
            spec_tmp = MetricSpec("tmp", tags, "instant", "USD", ["10-K", "10-Q"])
            cand = _filter_unit(cand, spec_tmp)
            if cand.empty:
                return None
            rec = pick_best_instant(cand, end=end_d, prefer_forms=["10-K", "10-Q"])
            if rec is None:
                return None
            try:
                return float(rec["val"])
            except Exception:
                return None

        def _hist_at(q: pd.Timestamp, metric: str) -> Optional[float]:
            if metric not in hq.columns:
                return None
            sub = hq[hq["quarter"] == q]
            if sub.empty:
                return None
            v = pd.to_numeric(sub.iloc[-1].get(metric), errors="coerce")
            return float(v) if pd.notna(v) else None

        def _add(sev: str, metric: str, msg: str, src: str) -> None:
            rows.append(
                {
                    "quarter": q_target.date(),
                    "metric": metric,
                    "severity": sev,
                    "message": msg,
                    "source": src,
                }
            )

        # Imported lazily to avoid widening the pipeline/module dependency cycle.
        from .pipeline import _pick_first_instant_tag, _pick_instant_tag

        # FY anchors + Q4 derivation sanity (Q4 = FY - 9M).
        dur_checks = [
            ("revenue", "Revenue"),
            ("net_income", "Net income"),
            ("cfo", "CFO"),
            ("capex", "Capex"),
        ]
        for metric_key, label in dur_checks:
            tags = list(spec_map.get(metric_key).tags if spec_map.get(metric_key) is not None else [])
            q4_tag = None
            q4_v = _hist_at(q_target, metric_key)
            if audit is not None and not audit.empty:
                aud_q4 = audit.copy()
                aud_q4["quarter"] = pd.to_datetime(aud_q4["quarter"], errors="coerce")
                aud_q4 = aud_q4[
                    (aud_q4["metric"] == metric_key)
                    & (aud_q4["quarter"] == q_target)
                ].copy()
                if not aud_q4.empty:
                    aud_q4["source_rank"] = aud_q4["source"].map(lambda s: 0 if str(s or "").lower() != "missing" else 1)
                    aud_q4 = aud_q4.sort_values(["source_rank"], ascending=[True])
                    q4_source = str(aud_q4.iloc[0].get("source") or "").lower()
                    q4_tag = str(aud_q4.iloc[0].get("tag") or "").strip()
                    if q4_source and q4_source not in {"derived_ytd_q4", "missing"}:
                        _add("info", "QA_Q4_Derivation", f"{label}: direct Q4 fact selected; FY-9M comparison skipped.", "history/audit")
                        continue
            fy_v = _pick_cf_duration(tags, q_target.date(), "FY", preferred_tag=q4_tag)
            y9_v = _pick_cf_duration(tags, q9.date(), "9M", preferred_tag=q4_tag)
            if metric_key == "capex":
                fy_v = abs(fy_v) if fy_v is not None else None
                y9_v = abs(y9_v) if y9_v is not None else None
                q4_v = abs(q4_v) if q4_v is not None else None
            if fy_v is None:
                _add("warn", "QA_FY_Anchor", f"{label}: FY fact missing for 2025-12-31.", "companyfacts")
                continue
            _add("info", "QA_FY_Anchor", f"{label}: FY fact {fy_v/1e6:,.1f}m found.", "companyfacts")
            if y9_v is None or q4_v is None:
                _add("warn", "QA_Q4_Derivation", f"{label}: missing 9M or Q4 value for FY-9M check.", "history/companyfacts")
                continue
            implied_q4 = fy_v - y9_v
            tol = max(5_000_000.0, 0.01 * max(1.0, abs(implied_q4)))
            diff = abs(q4_v - implied_q4)
            sev = "info" if diff <= tol else "warn"
            _add(
                sev,
                "QA_Q4_Derivation",
                f"{label}: Q4 {q4_v/1e6:,.1f}m vs (FY-9M) {implied_q4/1e6:,.1f}m (diff {diff/1e6:,.1f}m).",
                "history/companyfacts",
            )
            if metric_key in {"revenue", "cfo"} and q4_v < 0:
                _add("warn", "QA_Q4_Derivation", f"{label}: Q4 is negative ({q4_v/1e6:,.1f}m).", "history")

        # 12/31 instants should match direct companyfacts facts.
        cash_hist = _hist_at(q_target, "cash")
        cash_cf = _pick_cf_instant(list(spec_map.get("cash").tags if spec_map.get("cash") is not None else []), q_target.date())
        if cash_hist is None or cash_cf is None:
            _add("warn", "QA_FY_Anchor", "Cash@12/31 missing in history or companyfacts.", "history/companyfacts")
        else:
            diff = abs(cash_hist - cash_cf)
            tol = max(2_000_000.0, 0.001 * max(1.0, abs(cash_cf)))
            _add(
                "info" if diff <= tol else "warn",
                "QA_FY_Anchor",
                f"Cash@12/31 history {cash_hist/1e6:,.1f}m vs companyfacts {cash_cf/1e6:,.1f}m.",
                "history/companyfacts",
            )

        debt_hist = _hist_at(q_target, "total_debt")
        debt_pick = compute_total_debt_instant(df_all, end=q_target.date(), prefer_forms=["10-K", "10-Q"])
        debt_cf = float(debt_pick.value) if debt_pick is not None and debt_pick.value is not None else None
        if debt_hist is None or debt_cf is None:
            _add("warn", "QA_FY_Anchor", "Debt@12/31 missing in history or companyfacts.", "history/companyfacts")
        else:
            diff = abs(debt_hist - debt_cf)
            tol = max(5_000_000.0, 0.001 * max(1.0, abs(debt_cf)))
            _add(
                "info" if diff <= tol else "warn",
                "QA_FY_Anchor",
                f"Debt@12/31 history {debt_hist/1e6:,.1f}m vs companyfacts {debt_cf/1e6:,.1f}m.",
                "history/companyfacts",
            )

        # Double-count guard: when noncurrent + current exists, total debt must not equal LT + current.
        rec_non = _pick_first_instant_tag(
            df_all,
            end=q_target.date(),
            tags=["LongTermDebtNoncurrent", "LongTermDebtAndCapitalLeaseObligations"],
            prefer_forms=["10-K", "10-Q"],
        )
        rec_long = _pick_instant_tag(df_all, end=q_target.date(), tag="LongTermDebt", prefer_forms=["10-K", "10-Q"])
        rec_cur = _pick_first_instant_tag(
            df_all,
            end=q_target.date(),
            tags=["LongTermDebtCurrent", "DebtCurrent"],
            prefer_forms=["10-K", "10-Q"],
        )
        non_v = float(rec_non["val"]) if rec_non is not None and pd.notna(rec_non.get("val")) else None
        long_v = float(rec_long["val"]) if rec_long is not None and pd.notna(rec_long.get("val")) else None
        cur_v = float(rec_cur["val"]) if rec_cur is not None and pd.notna(rec_cur.get("val")) else None
        if debt_cf is not None and non_v is not None and cur_v is not None and long_v is not None:
            tol = max(5_000_000.0, 0.01 * max(1.0, abs(debt_cf)))
            bad_double = abs(debt_cf - (long_v + cur_v)) <= tol and abs(debt_cf - (non_v + cur_v)) > tol
            if bad_double:
                _add(
                    "fail",
                    "QA_Debt_DoubleCount",
                    "FAIL: total_debt matches LongTermDebt + current while noncurrent + current exists.",
                    "companyfacts",
                )
            else:
                _add(
                    "info",
                    "QA_Debt_DoubleCount",
                    "PASS: debt aggregation avoids LongTermDebt + current double-count path.",
                    "companyfacts",
                )

        if not rows:
            return info_df
        qa_df = pd.DataFrame(rows)
        return pd.concat([info_df, qa_df], ignore_index=True) if info_df is not None and not info_df.empty else qa_df

    def _append_key_metric_coverage_rows(info_df: pd.DataFrame) -> pd.DataFrame:
        if hist is None or hist.empty or "quarter" not in hist.columns:
            return info_df
        hcov = hist.copy()
        hcov["quarter"] = pd.to_datetime(hcov["quarter"], errors="coerce")
        hcov = hcov[hcov["quarter"].notna()].sort_values("quarter")
        if hcov.empty:
            return info_df
        if "fcf" not in hcov.columns and {"cfo", "capex"}.issubset(hcov.columns):
            hcov["fcf"] = pd.to_numeric(hcov["cfo"], errors="coerce") - pd.to_numeric(hcov["capex"], errors="coerce")
        key_metrics = ["cfo", "capex", "da", "ebitda", "fcf"]
        present_metrics = [m for m in key_metrics if m in hcov.columns]
        if not present_metrics:
            return info_df
        # Keep signal high: last 40 quarters and any quarter >=2015.
        htail = hcov.tail(40).copy()
        h2015 = hcov[hcov["quarter"].dt.year >= 2015].copy()
        check = (
            pd.concat([htail, h2015], ignore_index=True)
            .drop_duplicates(subset=["quarter"])
            .sort_values("quarter")
        )
        rows: List[Dict[str, Any]] = []
        for _, rr in check.iterrows():
            qd = pd.Timestamp(rr["quarter"]).date()
            missing = []
            for m in present_metrics:
                v = pd.to_numeric(rr.get(m), errors="coerce")
                if pd.isna(v):
                    missing.append(m)
            if missing:
                rows.append(
                    {
                        "quarter": qd,
                        "metric": "QA_Coverage",
                        "severity": "warn",
                        "message": f"Missing key metrics: {', '.join(missing)}",
                        "source": "history_coverage",
                    }
                )
            else:
                rows.append(
                    {
                        "quarter": qd,
                        "metric": "QA_Coverage",
                        "severity": "info",
                        "message": f"Key metrics present: {', '.join(present_metrics)}",
                        "source": "history_coverage",
                    }
                )
        if not rows:
            return info_df
        cov_df = pd.DataFrame(rows)
        return pd.concat([info_df, cov_df], ignore_index=True) if info_df is not None and not info_df.empty else cov_df

    info_log = _append_fy_q4_qa_rows(info_log)
    info_log = _append_key_metric_coverage_rows(info_log)
    if period_checks is not None and not period_checks.empty:
        fails = period_checks[period_checks["status"] == "fail"].copy()
        if not fails.empty:
            pr = fails.assign(
                severity="fail",
                source=fails["check"],
                message=fails["message"],
            )[["quarter", "metric", "severity", "message", "source"]]
            issues3 = pd.concat([issues3, pr], ignore_index=True)
    if not debt_tranches.empty and "scale_applied" in debt_tranches.columns:
        scaled = debt_tranches[debt_tranches["scale_applied"] != 1.0]
        if not scaled.empty:
            scale_rows = (
                scaled[["quarter", "scale_applied"]]
                .drop_duplicates()
                .assign(metric="debt_tranches", severity="warn",
                        message="Debt table scale inferred (thousands/millions). Review scaling.",
                        source="scale_inferred")
            )
            issues3 = pd.concat([issues3, scale_rows], ignore_index=True)
    if not non_gaap_needs.empty:
        issues3 = pd.concat([issues3, non_gaap_needs], ignore_index=True)

    if qa_checks is not None and not qa_checks.empty:
        qa_issues = qa_checks[
            ((qa_checks["status"].isin(["fail"])) | (qa_checks["check"] == "capex_negative"))
            & (qa_checks["check"] != "cash_identity")
        ].copy()
        if not qa_issues.empty:
            qa_rows = qa_issues.assign(
                severity=qa_issues["status"],
                source=qa_issues["check"],
                message=qa_issues["message"],
            )[["quarter", "metric", "severity", "message", "source"]]
            issues3 = pd.concat([issues3, qa_rows], ignore_index=True)

    needs_review = finalize_needs_review(
        concat_frames(issues1, issues2, issues3),
        review_quarters=config.qa_review_quarters,
    )

    debt_recon = issues2
    manifest_df = pd.DataFrame(getattr(sec, "manifest_rows", []))
    ocr_log_df = pd.DataFrame(getattr(sec, "ocr_log_rows", []))
    if ocr_log_df.empty:
        ocr_log_df = pd.DataFrame(columns=[
            "accn", "doc", "quarter", "purpose", "status",
            "image_files", "n_images", "text_len", "text_excerpt", "ocr_tokens",
            "report_date", "filing_date",
        ])
    else:
        for col in ["accn", "doc", "quarter", "purpose", "status", "image_files", "n_images", "text_len", "text_excerpt", "ocr_tokens", "report_date", "filing_date"]:
            if col not in ocr_log_df.columns:
                ocr_log_df[col] = None

    debt_schedule_key = (
        f"v2|max_quarters={config.max_quarters}|sub={_sub_recent_signature(sub, forms_prefix=('10-Q', '10-K'), max_rows=300)}"
    )
    debt_schedule_cached = _load_stage_cache("debt_schedule", debt_schedule_key)
    if isinstance(debt_schedule_cached, pd.DataFrame):
        debt_schedule = debt_schedule_cached
    else:
        debt_schedule = build_debt_schedule_tier2(sec, cik_int, sub, max_quarters=config.max_quarters, min_year=config.min_year)
        _save_stage_cache("debt_schedule", debt_schedule_key, debt_schedule)

    debt_profile, debt_tranches_latest, debt_maturity, debt_profile_qa, debt_profile_info = build_debt_profile(
        hist,
        df_all,
        debt_tranches,
        slides_debt=slides_debt,
        debt_schedule=debt_schedule,
    )
    debt_notes_key = (
        f"v1|max_docs=8|sub={_sub_recent_signature(sub, forms_prefix=('10-Q', '10-K', '8-K'), max_rows=300)}"
    )
    debt_credit_notes_cached = _load_stage_cache("debt_credit_notes", debt_notes_key)
    if isinstance(debt_credit_notes_cached, pd.DataFrame):
        debt_credit_notes = debt_credit_notes_cached
    else:
        debt_credit_notes = build_debt_credit_notes(sec, cik_int, sub, max_docs=8)
        _save_stage_cache("debt_credit_notes", debt_notes_key, debt_credit_notes)

    revolver_key = (
        f"v2|max_docs=80|lookback=7|sub={_sub_recent_signature(sub, forms_prefix=('10-Q', '10-K', '8-K'), max_rows=500)}"
    )
    revolver_df_cached = _load_stage_cache("revolver_df", revolver_key)
    if isinstance(revolver_df_cached, pd.DataFrame):
        revolver_df = revolver_df_cached
    else:
        revolver_df = build_revolver_availability(sec, cik_int, sub, max_docs=80, lookback_years=7)
        _save_stage_cache("revolver_df", revolver_key, revolver_df)
    rev_capacity_map, rev_capacity_meta = build_revolver_capacity_map(df_all, hist)
    revolver_history = build_revolver_history(
        revolver_df,
        hist,
        capacity_map=rev_capacity_map,
        capacity_meta=rev_capacity_meta,
        max_quarters=20,
    )
    local_main_revolver = build_local_main_revolver_history(
        base_dir,
        ticker=ticker,
        cache_root=Path(config.cache_dir) if config.cache_dir is not None else None,
        rebuild_doc_text_cache=config.rebuild_doc_text_cache,
        quiet_pdf_warnings=config.quiet_pdf_warnings,
    )
    if local_main_revolver is not None and not local_main_revolver.empty:
        local_main_revolver = local_main_revolver.copy()
        local_main_revolver["quarter"] = pd.to_datetime(local_main_revolver["quarter"], errors="coerce")
        if revolver_history is None or revolver_history.empty:
            revolver_history = local_main_revolver.copy()
        else:
            revolver_history = revolver_history.copy()
            revolver_history["quarter"] = pd.to_datetime(revolver_history["quarter"], errors="coerce")
            overlay_cols = [
                "revolver_commitment",
                "revolver_facility_size",
                "revolver_drawn",
                "revolver_letters_of_credit",
                "revolver_availability",
                "commitment_source_type",
                "facility_source_type",
                "drawn_source_type",
                "lc_source_type",
                "availability_source_type",
                "commitment_snippet",
                "drawn_snippet",
                "lc_snippet",
                "availability_snippet",
                "source_type",
                "source_snippet",
                "note",
            ]
            local_by_q = {
                pd.Timestamp(r["quarter"]).normalize(): r
                for _, r in local_main_revolver.dropna(subset=["quarter"]).iterrows()
            }
            existing_q = set()
            for idx, row in revolver_history.iterrows():
                qk = pd.Timestamp(row.get("quarter")).normalize() if pd.notna(row.get("quarter")) else None
                if qk is None or qk not in local_by_q:
                    continue
                existing_q.add(qk)
                local_row = local_by_q[qk]
                for col in overlay_cols:
                    if col in local_row.index:
                        local_val = local_row.get(col)
                        if pd.notna(local_val) and local_val != "":
                            revolver_history.at[idx, col] = local_val
                commit = pd.to_numeric(revolver_history.at[idx, "revolver_commitment"], errors="coerce")
                drawn = pd.to_numeric(revolver_history.at[idx, "revolver_drawn"], errors="coerce")
                avail = pd.to_numeric(revolver_history.at[idx, "revolver_availability"], errors="coerce")
                lc = pd.to_numeric(revolver_history.at[idx, "revolver_letters_of_credit"], errors="coerce")
                if pd.notna(commit) and pd.notna(drawn):
                    revolver_history.at[idx, "revolver_utilization"] = float(drawn) / float(commit) if float(commit) else None
                if pd.notna(commit) and pd.notna(drawn) and pd.notna(avail) and pd.isna(lc):
                    residual = float(commit) - float(drawn) - float(avail)
                    if residual >= -1_000_000.0:
                        revolver_history.at[idx, "revolver_letters_of_credit"] = max(residual, 0.0)
                        revolver_history.at[idx, "lc_source_type"] = "derived"
            missing_rows: List[Dict[str, Any]] = []
            for qk, local_row in local_by_q.items():
                if qk in existing_q:
                    continue
                new_row = {"quarter": qk}
                for col in overlay_cols:
                    if col in local_row.index:
                        new_row[col] = local_row.get(col)
                commit = pd.to_numeric(new_row.get("revolver_commitment"), errors="coerce")
                drawn = pd.to_numeric(new_row.get("revolver_drawn"), errors="coerce")
                avail = pd.to_numeric(new_row.get("revolver_availability"), errors="coerce")
                lc = pd.to_numeric(new_row.get("revolver_letters_of_credit"), errors="coerce")
                if pd.notna(commit) and pd.notna(drawn):
                    new_row["revolver_utilization"] = float(drawn) / float(commit) if float(commit) else None
                if pd.notna(commit) and pd.notna(drawn) and pd.notna(avail) and pd.isna(lc):
                    residual = float(commit) - float(drawn) - float(avail)
                    if residual >= -1_000_000.0:
                        new_row["revolver_letters_of_credit"] = max(residual, 0.0)
                        new_row["lc_source_type"] = "derived"
                missing_rows.append(new_row)
            if missing_rows:
                revolver_history = pd.concat([revolver_history, pd.DataFrame(missing_rows)], ignore_index=True, sort=False)
            revolver_history = revolver_history.sort_values("quarter").reset_index(drop=True)
    if str(ticker or "").upper() == "ANF":
        anf_abl_q = pd.Timestamp("2026-01-31")
        anf_abl_row = {
            "quarter": anf_abl_q,
            "revolver_commitment": 500_000_000.0,
            "revolver_facility_size": 500_000_000.0,
            "revolver_drawn": 0.0,
            "revolver_letters_of_credit": 454_000.0,
            "revolver_availability": 449_546_000.0,
            "revolver_utilization": 0.0,
            "commitment_source_type": "10-K debt note",
            "facility_source_type": "10-K debt note",
            "drawn_source_type": "10-K debt note",
            "lc_source_type": "10-K debt note",
            "availability_source_type": "10-K debt note",
            "source_type": "10-K debt note",
            "source_snippet": (
                "ABL Facility up to $500 million, matures August 2, 2029; no borrowings outstanding "
                "as of January 31, 2026; borrowing capacity available $449.546m."
            ),
            "note": "ANF latest ABL facility from FY2025 10-K.",
        }
        if revolver_history is None or revolver_history.empty:
            revolver_history = pd.DataFrame([anf_abl_row])
        else:
            revolver_history = revolver_history.copy()
            revolver_history["quarter"] = pd.to_datetime(revolver_history["quarter"], errors="coerce")
            mask = revolver_history["quarter"].dt.normalize().eq(anf_abl_q)
            if mask.any():
                idx = revolver_history.index[mask][-1]
                for col, val in anf_abl_row.items():
                    revolver_history.at[idx, col] = val
            else:
                revolver_history = pd.concat([revolver_history, pd.DataFrame([anf_abl_row])], ignore_index=True, sort=False)
            revolver_history = revolver_history.sort_values("quarter").reset_index(drop=True)
    debt_buckets, debt_bucket_qa = build_debt_buckets(debt_tranches_latest, hist, maturity_df=debt_maturity)
    try:
        if debt_buckets is not None and not debt_buckets.empty and debt_profile is not None and not debt_profile.empty:
            src = str(debt_buckets.iloc[0].get("Source") or "")
            if src == "scheduled_repayments_fallback":
                dpf = debt_profile.copy()
                if "metric" in dpf.columns and "value" in dpf.columns:
                    basis_vals = pd.to_numeric(
                        dpf.loc[dpf["metric"].astype(str).isin(["debt_long_term", "debt_principal_total"]), "value"],
                        errors="coerce",
                    ).dropna()
                    basis_val = float(basis_vals.iloc[-1]) if not basis_vals.empty else None
                    total_bucketed = pd.to_numeric(debt_buckets.iloc[0].get("Total_bucketed"), errors="coerce")
                    if basis_val not in (None, 0) and pd.notna(total_bucketed):
                        debt_buckets.loc[:, "Debt_long_term"] = basis_val
                        debt_buckets.loc[:, "Coverage_basis_metric"] = "debt_long_term"
                        debt_buckets.loc[:, "Coverage_basis_value"] = basis_val
                        debt_buckets.loc[:, "Bucket_coverage_pct"] = float(total_bucketed) / float(basis_val)
    except Exception:
        pass
    if debt_bucket_qa is not None and not debt_bucket_qa.empty:
        if qa_checks is None or qa_checks.empty:
            qa_checks = debt_bucket_qa
        else:
            qa_checks = pd.concat([qa_checks, debt_bucket_qa], ignore_index=True)
    if debt_profile_qa is not None and not debt_profile_qa.empty:
        if qa_checks is None or qa_checks.empty:
            qa_checks = debt_profile_qa
        else:
            qa_checks = pd.concat([qa_checks, debt_profile_qa], ignore_index=True)
    if debt_profile_info is not None and not debt_profile_info.empty:
        if info_log is None or info_log.empty:
            info_log = debt_profile_info
        else:
            info_log = pd.concat([info_log, debt_profile_info], ignore_index=True)

    earnings_release_candidates = [
        base_dir / "earnings_release",
        base_dir / "Earnings Release",
        base_dir / "Earnings Releases",
        base_dir / "press_release",
        base_dir / "Press Release",
    ]
    earnings_release_dir = next((p for p in earnings_release_candidates if p.exists() and p.is_dir()), None)
    earnings_release_sig = "none"
    if earnings_release_dir is not None:
        try:
            rel_rows: List[str] = []
            for fp in sorted([x for x in earnings_release_dir.glob("*.pdf") if x.is_file()], key=lambda x: x.name.lower())[:200]:
                st = fp.stat()
                rel_rows.append(f"{fp.name}:{int(st.st_size)}:{int(st.st_mtime)}")
            earnings_release_sig = hashlib.sha1("||".join(rel_rows).encode("utf-8", errors="ignore")).hexdigest()
        except Exception:
            earnings_release_sig = "err"

    # `doc_intel_bundle` is the bridge from raw document text into visible evidence
    # products such as Quarter_Notes_UI, promises, and promise-progress rows.
    # `doc_intel_bundle` is one of the most expensive stages because it turns raw
    # filing/local text into visible Quarter_Notes, promises, promise-progress, and
    # non-GAAP credibility evidence. The key therefore tracks both input content and
    # behavior-sensitive code signatures.
    doc_intel_key = "|".join(
        [
            DOC_INTEL_BEHAVIOR_VERSION,
            f"sub={_sub_recent_signature(sub, forms_prefix=('10-Q', '10-K', '8-K'), max_rows=500)}",
            f"hist={_df_quick_sig(hist, ['quarter', 'revenue', 'ebitda', 'fcf', 'debt', 'cash'])}",
            f"adj={_df_quick_sig(adj_metrics, ['quarter', 'adj_ebitda', 'adj_ebit', 'adj_eps'])}",
            f"revh={_df_quick_sig(revolver_history, ['quarter', 'revolver_commitment', 'revolver_drawn', 'revolver_availability'])}",
            f"db={_df_quick_sig(debt_buckets, ['quarter', 'maturity_year', 'amount_total'])}",
            f"er={earnings_release_sig}",
            f"max_docs={config.doc_intel_max_docs}",
            f"max_quarters={config.doc_intel_max_quarters}",
            "doc_text_cache=v2",
            f"code={_module_code_signature('doc_intel.py', 'quarter_notes.py')}",
            f"quiet_pdf={int(bool(config.quiet_pdf_warnings))}",
        ]
    )
    doc_intel_cached = _load_stage_cache("doc_intel_bundle", doc_intel_key)
    if isinstance(doc_intel_cached, dict):
        quarter_notes = doc_intel_cached.get("quarter_notes", pd.DataFrame())
        promises = doc_intel_cached.get("promises", pd.DataFrame())
        promise_progress = doc_intel_cached.get("promise_progress", pd.DataFrame())
        non_gaap_cred = doc_intel_cached.get("non_gaap_cred", pd.DataFrame())
    else:
        if config.use_cached_doc_intel_only:
            raise RuntimeError("doc_intel cache required but missing for --skip-doc-intel run.")
        with _timed_stage(stage_timings, "doc_intel_bundle", enabled=config.profile_timings):
            quarter_notes, promises, promise_progress, non_gaap_cred = build_doc_intel_outputs(
                sec=sec,
                cik_int=cik_int,
                submissions=sub,
                hist=hist,
                adj_metrics=adj_metrics,
                adj_breakdown=adj_breakdown,
                non_gaap_files=non_gaap_files,
                revolver_history=revolver_history,
                debt_buckets=debt_buckets,
                earnings_release_dir=earnings_release_dir,
                max_docs=config.doc_intel_max_docs,
                max_quarters=config.doc_intel_max_quarters,
                cache_dir=config.cache_dir,
                rebuild_doc_text_cache=config.rebuild_doc_text_cache,
                quiet_pdf_warnings=config.quiet_pdf_warnings,
                stage_timings=stage_timings,
                profile_timings=config.profile_timings,
            )
        _save_stage_cache(
            "doc_intel_bundle",
            doc_intel_key,
            {
                "quarter_notes": quarter_notes,
                "promises": promises,
                "promise_progress": promise_progress,
                "non_gaap_cred": non_gaap_cred,
            },
        )
    if tkr_u == "ANF":
        anf_source_notes = _build_anf_source_quarter_notes(hist=hist, base_dir=base_dir, config=config)
        if not anf_source_notes.empty:
            quarter_notes = pd.concat([quarter_notes, anf_source_notes], ignore_index=True, sort=False) if quarter_notes is not None and not quarter_notes.empty else anf_source_notes
        anf_progress = _build_anf_guidance_progress_rows(slides_guidance, hist=hist, adj_metrics=adj_metrics)
        if not anf_progress.empty:
            promise_progress = (
                pd.concat([promise_progress, anf_progress], ignore_index=True, sort=False)
                if promise_progress is not None and not promise_progress.empty
                else anf_progress
            )
    quarter_notes_qa = validate_quarter_notes(quarter_notes, hist)
    promise_qa_df = build_promise_qa_checks(promises, promise_progress)
    non_gaap_qa_df = build_non_gaap_cred_qa(non_gaap_cred)
    qa_checks = finalize_qa_checks(
        concat_frames(
            qa_checks,
            quarter_notes_qa,
            promise_qa_df,
            non_gaap_qa_df,
        ),
        review_quarters=config.qa_review_quarters,
    )

    # Company overview is cached independently because it is topic-aware summary text,
    # not just another generic dataframe side effect of the main pipeline.
    company_overview_key = "|".join(
        [
            COMPANY_OVERVIEW_BEHAVIOR_VERSION,
            f"sub={submissions_sig}",
            f"materials={local_material_sig}",
            f"ticker={str(ticker or '').upper()}",
            f"code={_module_code_signature('summary_overview.py')}",
        ]
    )
    company_overview_cached = _load_stage_cache("company_overview", company_overview_key)
    if isinstance(company_overview_cached, dict):
        company_overview = company_overview_cached
    else:
        try:
            with _timed_stage(stage_timings, "company_overview", enabled=config.profile_timings):
                company_overview = build_company_overview(sec, cik_int, sub, ticker=ticker)
            _save_stage_cache("company_overview", company_overview_key, company_overview)
        except Exception as exc:
            # Safe blank behavior: when summary evidence cannot be resolved, return
            # explicit `N/A` placeholders instead of guessing narrative text that
            # could contaminate visible workbook surfaces.
            err = f"{type(exc).__name__}: {exc}"
            company_overview = {
                "what_it_does": "N/A",
                "what_it_does_source": f"Source: N/A ({err})",
                "current_strategic_context": "N/A",
                "current_strategic_context_source": f"Source: N/A ({err})",
                "key_advantage": "N/A",
                "key_advantage_source": f"Source: N/A ({err})",
                "revenue_streams": [],
                "revenue_streams_source": f"Source: N/A ({err})",
                "asof_fy_end": None,
            }
    if tkr_u == "ANF":
        company_overview = _apply_anf_company_overview_overrides(
            company_overview,
            slides_segments=slides_segments,
        )

    if config.profile_timings and stage_timings:
        summary = " | ".join(f"{k}={v:.2f}s" for k, v in sorted(stage_timings.items(), key=lambda kv: (-kv[1], kv[0])))
        print(f"[run_pipeline timing] {summary}", flush=True)

    # Hard regression gate: fail fast on key invariants.
    from .pipeline import _regression_gate

    _regression_gate(
        hist,
        audit,
        df_all,
        ticker=tkr_u,
        cache_dir=config.cache_dir,
        debug=config.debug_regression_gate,
        allow_fail=config.allow_regression_gate_fail,
    )
    return PipelineArtifacts(
        hist=hist,
        audit=audit,
        debt_tranches=debt_tranches,
        debt_recon=debt_recon,
        adj_metrics=adj_metrics,
        adj_breakdown=adj_breakdown,
        non_gaap_files=non_gaap_files,
        adj_metrics_relaxed=adj_metrics_relaxed,
        adj_breakdown_relaxed=adj_breakdown_relaxed,
        non_gaap_files_relaxed=non_gaap_files_relaxed,
        needs_review=needs_review,
        info_log=info_log,
        tag_coverage=tag_coverage,
        period_checks=period_checks,
        qa_checks=qa_checks,
        bridge_q=bridge_q,
        manifest_df=manifest_df,
        ocr_log=ocr_log_df,
        qfd_preview=qfd_preview,
        qfd_unused=qfd_unused,
        debt_profile=debt_profile,
        debt_tranches_latest=debt_tranches_latest,
        debt_maturity=debt_maturity,
        debt_credit_notes=debt_credit_notes,
        revolver_df=revolver_df,
        revolver_history=revolver_history,
        debt_buckets=debt_buckets,
        slides_segments=slides_segments,
        slides_debt=slides_debt,
        slides_guidance=slides_guidance,
        quarter_notes=quarter_notes,
        promises=promises,
        promise_progress=promise_progress,
        non_gaap_cred=non_gaap_cred,
        company_overview=company_overview,
        guidance_raw=slides_guidance_raw,
        stage_timings=stage_timings,
    )
