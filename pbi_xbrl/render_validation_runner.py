"""Render and openpyxl style validation for generated stock-model workbooks.

The Excel COM renderer is intentionally optional. CI and many desktop sessions
do not have a working Excel automation host, so render validation degrades to a
clear "render skipped" result while keeping openpyxl layout checks active.
"""
from __future__ import annotations

import argparse
import csv
import gc
import hashlib
import json
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Optional, Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from .new_engine_excel import (
    _co_initialize,
    _co_uninitialize,
    _dispatch_excel,
    _excel_process_id,
    _wait_for_owned_process_exit,
)
from .path_config import resolve_stock_model_paths
from .workbook_modules import DEFAULT_MODULE_MANIFEST, load_workbook_module_manifest


RENDER_RANGES: Dict[str, str] = {
    "Valuation": "A1:AC90",
    "{ticker}_Investment_Case": "A1:J160",
    "Promise_Progress_UI": "A1:L180",
    "Quarter_Notes_UI": "A1:O220",
    "Operating_Drivers": "A1:Q140",
    "Needs_Review": "A1:J80",
}

_RENDER_OPERATION_NAMES: tuple[str, ...] = (
    "open",
    "activate_sheet",
    "resolve_range",
    "copy_picture",
    "scratch_chart_add",
    "paste",
    "export",
    "nonblank_check",
    "cleanup",
)

USER_FACING_STYLE_SHEETS: Sequence[str] = (
    "Valuation",
    "{ticker}_Investment_Case",
    "Promise_Progress_UI",
    "Quarter_Notes_UI",
)

TITLE_FILLS = {"005B9BD5", "004472C4", "006FA8DC"}
SUBHEADER_FILLS = {"00DDEBF7", "00EAF3F8", "00BDD7EE", "009DCEF0"}
BODY_FILLS = {"00F7FBFF", "00EDF4FB", "00FFFFFF", "00F2F2F2", "00EAF3F8"}
MANUAL_FILL_SUFFIX = "FFF2CC"


class RenderUnavailableError(RuntimeError):
    """Raised when optional Excel COM rendering cannot be loaded on this host."""


@dataclass(frozen=True)
class StyleValidationIssue:
    severity: str
    sheet: str
    cell: str
    message: str

    def to_dict(self) -> Dict[str, str]:
        return {
            "severity": self.severity,
            "sheet": self.sheet,
            "cell": self.cell,
            "message": self.message,
        }


@dataclass
class OpenpyxlLayoutReport:
    ticker: str
    workbook_path: Path
    checked_sheets: int = 0
    max_row_height: float = 0.0
    issues: list[StyleValidationIssue] = field(default_factory=list)

    @property
    def overall(self) -> str:
        return "FAIL" if any(issue.severity == "error" for issue in self.issues) else "PASS"

    def to_dict(self) -> Dict[str, Any]:
        return {
            "ticker": self.ticker,
            "workbook_path": str(self.workbook_path),
            "checked_sheets": self.checked_sheets,
            "max_row_height": self.max_row_height,
            "overall": self.overall,
            "issues": [issue.to_dict() for issue in self.issues],
        }


@dataclass(frozen=True)
class RenderTarget:
    sheet: str
    range_ref: str
    source: str
    block_id: str = ""
    header_row: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "sheet": self.sheet,
            "range": self.range_ref,
            "source": self.source,
            "block_id": self.block_id,
            "header_row": self.header_row,
        }


@dataclass
class RenderedRangeResult:
    ticker: str
    sheet: str
    range_ref: str
    image_path: Optional[Path] = None
    status: str = "pending"
    message: str = ""
    failing_operation: str = ""
    cleanup_failure: dict[str, Any] | None = None
    operations: list[dict[str, Any]] = field(default_factory=list)
    source_sha256_before: str = ""
    source_sha256_after: str = ""
    source_bytes_unchanged: bool | None = None
    owned_excel_process_id: int | None = None
    owned_excel_process_cleanup: str = "pending"
    owned_excel_process_forced_termination: bool = False

    def to_dict(self) -> Dict[str, Any]:
        return {
            "ticker": self.ticker,
            "sheet": self.sheet,
            "range": self.range_ref,
            "image_path": str(self.image_path) if self.image_path else "",
            "status": self.status,
            "message": self.message,
            "failing_operation": self.failing_operation,
            "cleanup_failure": self.cleanup_failure,
            "operations": self.operations,
            "source_sha256_before": self.source_sha256_before,
            "source_sha256_after": self.source_sha256_after,
            "source_bytes_unchanged": self.source_bytes_unchanged,
            "owned_excel_process_id": self.owned_excel_process_id,
            "owned_excel_process_cleanup": self.owned_excel_process_cleanup,
            "owned_excel_process_forced_termination": self.owned_excel_process_forced_termination,
        }


@dataclass
class RenderValidationReport:
    output_dir: Path
    render_status: str
    skip_reason: str = ""
    style_reports: Dict[str, OpenpyxlLayoutReport] = field(default_factory=dict)
    rendered_ranges: list[RenderedRangeResult] = field(default_factory=list)
    elapsed_seconds: float = 0.0

    @property
    def overall(self) -> str:
        if any(report.overall == "FAIL" for report in self.style_reports.values()):
            return "FAIL"
        if any(result.status == "fail" for result in self.rendered_ranges):
            return "FAIL"
        if self.render_status == "fail":
            return "FAIL"
        return "SKIP_RENDER" if self.render_status == "skipped" else "PASS"

    def to_dict(self) -> Dict[str, Any]:
        return {
            "output_dir": str(self.output_dir),
            "render_status": self.render_status,
            "skip_reason": self.skip_reason,
            "overall": self.overall,
            "elapsed_seconds": self.elapsed_seconds,
            "style_reports": {ticker: report.to_dict() for ticker, report in self.style_reports.items()},
            "rendered_ranges": [result.to_dict() for result in self.rendered_ranges],
        }

    def to_summary_rows(self) -> list[Dict[str, Any]]:
        rows: list[Dict[str, Any]] = []
        for ticker, report in self.style_reports.items():
            render_failures = [
                result
                for result in self.rendered_ranges
                if result.ticker == ticker and result.status == "fail"
            ]
            rows.append(
                {
                    "Ticker": ticker,
                    "Style": report.overall,
                    "Style issues": len([issue for issue in report.issues if issue.severity == "error"]),
                    "Max row height": report.max_row_height,
                    "Render": self.render_status,
                    "Render failures": len(render_failures),
                    "Overall": "FAIL" if report.overall == "FAIL" or render_failures else self.overall,
                }
            )
        return rows


def _text(value: Any) -> str:
    return str(value or "").strip()


def _rgb(cell: Cell) -> str:
    return _text(cell.fill.fgColor.rgb).upper()


def _effective_cell(ws: Worksheet, row: int, col: int) -> Cell:
    coord = ws.cell(row, col).coordinate
    for merged in ws.merged_cells.ranges:
        if coord in merged:
            return ws.cell(merged.min_row, merged.min_col)
    return ws.cell(row, col)


def _effective_rgb(ws: Worksheet, row: int, col: int) -> str:
    return _rgb(_effective_cell(ws, row, col))


def _effective_border_present(ws: Worksheet, row: int, col: int) -> bool:
    cell = _effective_cell(ws, row, col)
    return any(
        _text(getattr(side, "style", ""))
        for side in (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom)
    )


def _effective_wrap(ws: Worksheet, row: int, col: int) -> bool:
    return bool(_effective_cell(ws, row, col).alignment.wrap_text)


def _row_has_merge(ws: Worksheet, row: int, min_col: int, max_col: int) -> bool:
    return any(
        merged.min_row == row
        and merged.max_row == row
        and merged.min_col <= min_col
        and merged.max_col >= max_col
        for merged in ws.merged_cells.ranges
    )


def _sheet_name(template: str, ticker: str) -> str:
    return template.format(ticker=ticker)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _manifest_render_targets(wb: Any, module_payload: Mapping[str, Any]) -> list[RenderTarget]:
    targets: list[RenderTarget] = []
    seen: set[tuple[str, str]] = set()
    for module in module_payload.get("modules") or []:
        sheet_contracts = {
            str(row.get("sheet") or ""): row
            for row in module.get("sheets") or []
            if isinstance(row, Mapping)
        }
        for block in module.get("visible_blocks") or []:
            if not isinstance(block, Mapping):
                continue
            sheet_name = str(block.get("sheet") or "")
            range_ref = str(block.get("target") or "")
            contract = sheet_contracts.get(sheet_name)
            if not contract or str(contract.get("default_state") or "") not in {"hidden", "veryHidden"}:
                continue
            if sheet_name not in wb.sheetnames or wb[sheet_name].sheet_state != "visible":
                continue
            key = (sheet_name, range_ref)
            if key in seen:
                continue
            seen.add(key)
            targets.append(
                RenderTarget(
                    sheet=sheet_name,
                    range_ref=range_ref,
                    source="module_manifest_visible_block",
                    block_id=str(block.get("block_id") or ""),
                    header_row=int(contract.get("header_row") or 0),
                )
            )
    return targets


def discover_render_targets(
    workbook_path: Path | str,
    ticker: str,
    *,
    module_manifest_path: Path | str = DEFAULT_MODULE_MANIFEST,
) -> tuple[RenderTarget, ...]:
    """Return core and conditionally visible product targets in deterministic order."""

    ticker_txt = str(ticker or "").strip().upper()
    wb = load_workbook(Path(workbook_path), data_only=False, read_only=False)
    try:
        targets = [
            RenderTarget(
                sheet=_sheet_name(sheet_template, ticker_txt),
                range_ref=range_ref,
                source="core_render_contract",
            )
            for sheet_template, range_ref in RENDER_RANGES.items()
        ]
        module_payload = load_workbook_module_manifest(module_manifest_path)
        existing = {(target.sheet, target.range_ref) for target in targets}
        targets.extend(
            target
            for target in _manifest_render_targets(wb, module_payload)
            if (target.sheet, target.range_ref) not in existing
        )
        return tuple(targets)
    finally:
        wb.close()


def _nonblank_count(ws: Worksheet, min_col: int, min_row: int, max_col: int, max_row: int) -> int:
    count = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if _text(cell.value):
                count += 1
    return count


def _add(report: OpenpyxlLayoutReport, severity: str, sheet: str, cell: str, message: str) -> None:
    report.issues.append(StyleValidationIssue(severity=severity, sheet=sheet, cell=cell, message=message))


def _validate_render_surfaces(
    wb: Any,
    report: OpenpyxlLayoutReport,
    targets: Sequence[RenderTarget],
) -> None:
    for target in targets:
        sheet_name = target.sheet
        range_ref = target.range_ref
        if sheet_name not in wb.sheetnames:
            _add(report, "error", sheet_name, "A1", "required render surface sheet is missing")
            continue
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            _add(report, "error", sheet_name, "A1", "major user-facing sheet is hidden")
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        if _nonblank_count(ws, min_col, min_row, max_col, min(max_row, int(ws.max_row or 0))) == 0:
            _add(report, "error", sheet_name, range_ref, "render range is fully blank")
        if target.source == "core_render_contract" and (max_col - min_col < 4 or max_row - min_row < 20):
            _add(report, "error", sheet_name, range_ref, "render range dimensions are unexpectedly small")
        if target.source == "module_manifest_visible_block":
            first_body_row = max(min_row, target.header_row + 1)
            body_nonblank = _nonblank_count(
                ws,
                min_col,
                first_body_row,
                max_col,
                min(max_row, int(ws.max_row or 0)),
            )
            if body_nonblank == 0:
                _add(report, "error", sheet_name, range_ref, "visible conditional product block is header-only")


def _validate_title_row(ws: Worksheet, report: OpenpyxlLayoutReport, sheet_name: str, expected_width: int) -> None:
    title = _text(ws.cell(1, 1).value)
    fill = _effective_rgb(ws, 1, 1)
    if not title:
        _add(report, "error", sheet_name, "A1", "missing title row text")
    if not (fill in TITLE_FILLS or fill.endswith(("5B9BD5", "4472C4", "6FA8DC"))):
        _add(report, "error", sheet_name, "A1", "missing title/header styling")
    if sheet_name in {"Promise_Progress_UI", "Quarter_Notes_UI"} and not _row_has_merge(ws, 1, 1, expected_width):
        _add(report, "error", sheet_name, "A1", "top title row is not merged across expected UI width")


def _validate_manual_inputs(ws: Worksheet, report: OpenpyxlLayoutReport, sheet_name: str) -> None:
    manual_header = None
    for rr in range(1, int(ws.max_row or 0) + 1):
        if _text(ws.cell(rr, 1).value) == "Manual Market / Scenario Inputs":
            manual_header = rr
            break
    if manual_header is None:
        return
    next_section = int(ws.max_row or 0) + 1
    for rr in range(manual_header + 1, int(ws.max_row or 0) + 1):
        first = _text(ws.cell(rr, 1).value)
        fill = _effective_rgb(ws, rr, 1)
        if rr > manual_header + 1 and first and fill.endswith(("5B9BD5", "4472C4", "6FA8DC")):
            next_section = rr
            break
    manual_cells_checked = 0
    for rr in range(manual_header + 2, next_section):
        if not _text(ws.cell(rr, 1).value):
            continue
        manual_cells_checked += 1
        fill = _rgb(ws.cell(rr, 6))
        if not fill.endswith(MANUAL_FILL_SUFFIX):
            _add(report, "error", sheet_name, f"F{rr}", "manual override cell is not yellow/editable-looking")
    if manual_cells_checked == 0:
        _add(report, "error", sheet_name, f"A{manual_header}", "manual input section has no input rows")


def _validate_ui_table_fill_and_borders(
    ws: Worksheet,
    report: OpenpyxlLayoutReport,
    sheet_name: str,
    *,
    width: int,
    max_height: float,
    narrative: bool = False,
) -> None:
    for rr in range(1, int(ws.max_row or 0) + 1):
        height = float(ws.row_dimensions[rr].height or 0.0)
        report.max_row_height = max(report.max_row_height, height)
        if height > max_height:
            _add(report, "error", sheet_name, str(rr), f"row height {height} exceeds validation band {max_height}")

        row_values = [_text(ws.cell(rr, cc).value) for cc in range(1, min(width, int(ws.max_column or 0)) + 1)]
        if not any(row_values):
            continue
        first = row_values[0]
        if first.endswith(" - Quarter Notes") or first.endswith("revisions") or first.endswith("guidance progression") or first.endswith("open guidance"):
            if _effective_rgb(ws, rr, 1) in {"", "00000000"}:
                _add(report, "error", sheet_name, f"A{rr}", "section header lacks fill")
            continue
        if first in {"Metric", "Theme", "Promise / guidance item", "Driver", "Category", "Input"}:
            for cc in range(1, min(width, int(ws.max_column or 0)) + 1):
                if _effective_rgb(ws, rr, cc) in {"", "00000000"}:
                    _add(report, "error", sheet_name, f"{ws.cell(rr, cc).coordinate}", "table header fill does not cover expected width")
            continue
        if sheet_name in {"Promise_Progress_UI", "Quarter_Notes_UI"}:
            for cc in range(1, min(width, int(ws.max_column or 0)) + 1):
                fill = _effective_rgb(ws, rr, cc)
                if fill in {"", "00000000"}:
                    _add(report, "error", sheet_name, ws.cell(rr, cc).coordinate, "zebra/body fill gap")
                if not _effective_border_present(ws, rr, cc):
                    _add(report, "error", sheet_name, ws.cell(rr, cc).coordinate, "missing thin border")
        if narrative and any(len(value) > 70 for value in row_values[1:]) and not any(_effective_wrap(ws, rr, cc) for cc in range(2, min(width, int(ws.max_column or 0)) + 1)):
            _add(report, "error", sheet_name, f"B{rr}", "long narrative row is not wrapped")


def validate_openpyxl_layout(
    workbook_path: Path | str,
    ticker: str,
    *,
    module_manifest_path: Path | str = DEFAULT_MODULE_MANIFEST,
    render_targets: Sequence[RenderTarget] | None = None,
) -> OpenpyxlLayoutReport:
    workbook_path = Path(workbook_path)
    ticker = str(ticker or "").strip().upper()
    report = OpenpyxlLayoutReport(ticker=ticker, workbook_path=workbook_path)
    targets = tuple(render_targets or discover_render_targets(
        workbook_path,
        ticker,
        module_manifest_path=module_manifest_path,
    ))
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        _validate_render_surfaces(wb, report, targets)
        for sheet_template in USER_FACING_STYLE_SHEETS:
            sheet_name = _sheet_name(sheet_template, ticker)
            if sheet_name not in wb.sheetnames:
                _add(report, "error", sheet_name, "A1", "required user-facing style sheet is missing")
                continue
            ws = wb[sheet_name]
            report.checked_sheets += 1
            expected_width = 15 if sheet_name == "Quarter_Notes_UI" else (10 if sheet_name != "Valuation" else 29)
            if sheet_name != "Valuation":
                _validate_title_row(ws, report, sheet_name, expected_width=expected_width)
            if sheet_name.endswith("_Investment_Case"):
                _validate_manual_inputs(ws, report, sheet_name)
            if sheet_name == "Valuation":
                for rr in range(1, int(ws.max_row or 0) + 1):
                    report.max_row_height = max(report.max_row_height, float(ws.row_dimensions[rr].height or 0.0))
                    if float(ws.row_dimensions[rr].height or 0.0) > 120.0:
                        _add(report, "error", sheet_name, str(rr), "row height exceeds validation band 120.0")
            else:
                _validate_ui_table_fill_and_borders(
                    ws,
                    report,
                    sheet_name,
                    width=expected_width,
                    max_height=95.0 if sheet_name == "Quarter_Notes_UI" else 120.0,
                    narrative=sheet_name == "Quarter_Notes_UI",
                )
    finally:
        wb.close()
    return report


def _image_nonblank(path: Path) -> tuple[bool, str]:
    if not path.exists() or path.stat().st_size < 200:
        return False, "image file missing or too small"
    try:
        from PIL import Image, ImageStat  # type: ignore

        with Image.open(path) as img:
            extrema = ImageStat.Stat(img.convert("L")).extrema[0]
            if extrema[0] == extrema[1]:
                return False, "image appears single-color/blank"
    except Exception:
        return True, "Pillow unavailable; file-size check only"
    return True, "nonblank"


def _com_error_details(exc: BaseException) -> tuple[int | None, str]:
    raw_hresult = getattr(exc, "hresult", None)
    if raw_hresult is None and getattr(exc, "args", ()):
        candidate = exc.args[0]
        raw_hresult = candidate if isinstance(candidate, int) else None
    detail = str(exc)
    excepinfo = getattr(exc, "excepinfo", None)
    if excepinfo:
        detail = f"{detail}; excepinfo={excepinfo!r}"
    return (int(raw_hresult) if raw_hresult is not None else None), detail


def _pending_operations() -> list[dict[str, Any]]:
    return [
        {
            "operation": name,
            "status": "pending",
            "attempts": 0,
            "elapsed_seconds": 0.0,
            "hresult": None,
            "message": "",
        }
        for name in _RENDER_OPERATION_NAMES
    ]


def _record_operation(
    result: RenderedRangeResult,
    operation: str,
    *,
    status: str,
    started: float,
    attempts: int = 1,
    exc: BaseException | None = None,
    message: str = "",
) -> None:
    row = next(item for item in result.operations if item["operation"] == operation)
    hresult, exception_message = _com_error_details(exc) if exc is not None else (None, "")
    row.update(
        {
            "status": status,
            "attempts": attempts,
            "elapsed_seconds": round(time.perf_counter() - started, 3),
            "hresult": hresult,
            "message": exception_message or message,
        }
    )
    if status == "fail" and not result.failing_operation:
        result.failing_operation = operation


def _record_cleanup_failures(
    result: RenderedRangeResult,
    failures: Sequence[tuple[str, BaseException]],
    *,
    started: float,
) -> None:
    if not failures:
        return
    primary_operation = result.failing_operation
    events = []
    for operation, exc in failures:
        hresult, message = _com_error_details(exc)
        events.append({"operation": operation, "hresult": hresult, "message": message})
    existing_events = list((result.cleanup_failure or {}).get("events") or [])
    all_events = [*existing_events, *events]
    detail = "; ".join(
        f"{event['operation']}={event['message']}" for event in all_events
    )
    result.cleanup_failure = {
        "hresult": next(
            (event["hresult"] for event in all_events if event["hresult"] is not None),
            None,
        ),
        "message": detail,
        "events": all_events,
    }
    _record_operation(
        result,
        "cleanup",
        status="fail",
        started=started,
        attempts=len(all_events),
        message=detail,
    )
    cleanup_row = next(item for item in result.operations if item["operation"] == "cleanup")
    cleanup_row["hresult"] = result.cleanup_failure["hresult"]
    if primary_operation:
        result.failing_operation = primary_operation
    result.status = "fail"
    if not primary_operation:
        result.message = detail


def _delete_owned_render_image(path: Path) -> None:
    path.unlink()


def _copy_picture_with_retry(rng: Any, result: RenderedRangeResult, *, attempts: int = 4) -> None:
    import pythoncom

    started = time.perf_counter()
    last_error: BaseException | None = None
    for attempt in range(1, attempts + 1):
        try:
            rng.CopyPicture(Appearance=1, Format=2)
            pythoncom.PumpWaitingMessages()
            _record_operation(
                result,
                "copy_picture",
                status="pass",
                started=started,
                attempts=attempt,
                message="range copied as enhanced metafile",
            )
            return
        except Exception as exc:  # pragma: no cover - COM timing is environment-dependent
            last_error = exc
            pythoncom.PumpWaitingMessages()
            time.sleep(0.2 * attempt)
    assert last_error is not None
    _record_operation(
        result,
        "copy_picture",
        status="fail",
        started=started,
        attempts=attempts,
        exc=last_error,
    )
    raise last_error


def _render_ranges_with_excel_com(
    workbooks: Mapping[str, Path],
    output_dir: Path,
    targets_by_ticker: Mapping[str, Sequence[RenderTarget]],
) -> list[RenderedRangeResult]:
    results: list[RenderedRangeResult] = []
    excel: Any = None
    source_wb: Any = None
    scratch_wb: Any = None
    scratch_ws: Any = None
    process_id: int | None = None
    initialized = False
    cleanup_result: dict[str, Any] = {"status": "pending", "forced_termination": False}
    primary_error: BaseException | None = None
    try:  # pragma: no cover - environment-dependent
        try:
            _co_initialize()
        except (ImportError, ModuleNotFoundError) as exc:
            raise RenderUnavailableError(f"Excel COM support is unavailable: {exc}") from exc
        initialized = True
        try:
            excel = _dispatch_excel()
        except (ImportError, ModuleNotFoundError) as exc:
            raise RenderUnavailableError(f"Excel COM support is unavailable: {exc}") from exc
        process_id = _excel_process_id(excel)
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        scratch_wb = excel.Workbooks.Add()
        scratch_ws = scratch_wb.Worksheets(1)
        scratch_ws.Unprotect()
        for ticker, workbook_path in workbooks.items():
            source_path = Path(workbook_path).resolve()
            source_hash_before = _sha256(source_path)
            open_started = time.perf_counter()
            targets = tuple(targets_by_ticker[ticker])
            try:
                source_wb = excel.Workbooks.Open(
                    str(source_path),
                    UpdateLinks=0,
                    ReadOnly=True,
                    IgnoreReadOnlyRecommended=True,
                    AddToMru=False,
                    CorruptLoad=0,
                )
            except Exception as exc:
                for target in targets:
                    result = RenderedRangeResult(
                        ticker=ticker,
                        sheet=target.sheet,
                        range_ref=target.range_ref,
                        operations=_pending_operations(),
                        source_sha256_before=source_hash_before,
                        owned_excel_process_id=process_id,
                    )
                    _record_operation(result, "open", status="fail", started=open_started, exc=exc)
                    result.status = "fail"
                    result.message = result.operations[0]["message"]
                    results.append(result)
                continue
            try:
                for target in targets:
                    result = RenderedRangeResult(
                        ticker=ticker,
                        sheet=target.sheet,
                        range_ref=target.range_ref,
                        operations=_pending_operations(),
                        source_sha256_before=source_hash_before,
                        owned_excel_process_id=process_id,
                    )
                    _record_operation(
                        result,
                        "open",
                        status="pass",
                        started=open_started,
                        message="source workbook opened read-only with links disabled",
                    )
                    image_path = output_dir / ticker / f"{target.sheet}_{target.range_ref.replace(':', '_')}.png"
                    image_path.parent.mkdir(parents=True, exist_ok=True)
                    result.image_path = image_path
                    chart_obj: Any = None
                    try:
                        started = time.perf_counter()
                        source_wb.Activate()
                        ws = source_wb.Worksheets(target.sheet)
                        if int(ws.Visible) not in {-1, 1}:
                            raise RuntimeError(f"Render target worksheet is not visible: {target.sheet}")
                        ws.Activate()
                        _record_operation(
                            result,
                            "activate_sheet",
                            status="pass",
                            started=started,
                            message="source workbook and visible worksheet activated",
                        )

                        started = time.perf_counter()
                        rng = ws.Range(target.range_ref)
                        rng.Select()
                        _record_operation(
                            result,
                            "resolve_range",
                            status="pass",
                            started=started,
                            message="bounded source range resolved and selected",
                        )

                        _copy_picture_with_retry(rng, result)

                        started = time.perf_counter()
                        scratch_wb.Activate()
                        scratch_ws.Activate()
                        chart_obj = scratch_ws.ChartObjects().Add(
                            0.0,
                            0.0,
                            max(1.0, float(rng.Width)),
                            max(1.0, float(rng.Height)),
                        )
                        _record_operation(
                            result,
                            "scratch_chart_add",
                            status="pass",
                            started=started,
                            message="owned scratch chart created outside source workbook",
                        )

                        started = time.perf_counter()
                        chart_obj.Chart.Paste()
                        _record_operation(result, "paste", status="pass", started=started)

                        started = time.perf_counter()
                        exported = bool(chart_obj.Chart.Export(str(image_path)))
                        if not exported:
                            raise RuntimeError("Excel Chart.Export returned false.")
                        _record_operation(result, "export", status="pass", started=started)

                        started = time.perf_counter()
                        ok, message = _image_nonblank(image_path)
                        _record_operation(
                            result,
                            "nonblank_check",
                            status="pass" if ok else "fail",
                            started=started,
                            message=message,
                        )
                        result.status = "pass" if ok else "fail"
                        result.message = message
                    except Exception as exc:
                        if not result.failing_operation:
                            pending = next(
                                item["operation"]
                                for item in result.operations
                                if item["status"] == "pending" and item["operation"] != "cleanup"
                            )
                            _record_operation(result, pending, status="fail", started=time.perf_counter(), exc=exc)
                        result.status = "fail"
                        result.message = next(
                            item["message"]
                            for item in result.operations
                            if item["operation"] == result.failing_operation
                        )
                    finally:
                        cleanup_started = time.perf_counter()
                        cleanup_failures: list[tuple[str, BaseException]] = []
                        try:
                            if chart_obj is not None:
                                chart_obj.Delete()
                        except Exception as exc:
                            cleanup_failures.append(("scratch_chart_delete", exc))
                        try:
                            excel.CutCopyMode = False
                        except Exception as exc:
                            cleanup_failures.append(("clipboard_clear", exc))
                        chart_obj = None
                        rng = None
                        ws = None
                        if result.status == "fail" and image_path.exists():
                            try:
                                _delete_owned_render_image(image_path)
                            except Exception as exc:
                                cleanup_failures.append(("temporary_file_delete", exc))
                        if cleanup_failures:
                            _record_cleanup_failures(
                                result,
                                cleanup_failures,
                                started=cleanup_started,
                            )
                        else:
                            _record_operation(
                                result,
                                "cleanup",
                                status="pass",
                                started=cleanup_started,
                                message="scratch chart, clipboard and owned image state released",
                            )
                    results.append(result)
            finally:
                source_wb.Close(SaveChanges=False)
                source_wb = None
                source_hash_after = _sha256(source_path)
                unchanged = source_hash_after == source_hash_before
                for result in results:
                    if result.ticker == ticker:
                        result.source_sha256_after = source_hash_after
                        result.source_bytes_unchanged = unchanged
                        if not unchanged:
                            result.status = "fail"
                            result.message = "source workbook bytes changed during read-only rendering"
                            if not result.failing_operation:
                                result.failing_operation = "cleanup"
                if not unchanged:
                    raise RuntimeError(f"Source workbook changed during read-only visual rendering: {source_path}")
    except Exception as exc:
        primary_error = exc
        raise
    finally:
        cleanup_started = time.perf_counter()
        cleanup_failures: list[tuple[str, BaseException]] = []
        if source_wb is not None:
            try:
                source_wb.Close(SaveChanges=False)
            except Exception as exc:
                cleanup_failures.append(("source_workbook_close", exc))
        if scratch_wb is not None:
            try:
                scratch_wb.Close(SaveChanges=False)
            except Exception as exc:
                cleanup_failures.append(("scratch_workbook_close", exc))
        source_wb = None
        scratch_ws = None
        scratch_wb = None
        if excel is not None:
            try:
                excel.Quit()
            except Exception as exc:
                cleanup_failures.append(("excel_quit", exc))
        excel = None
        try:
            gc.collect()
        except Exception as exc:
            cleanup_failures.append(("com_proxy_release", exc))
        if initialized:
            try:
                _co_uninitialize()
            except Exception as exc:
                cleanup_failures.append(("co_uninitialize", exc))
        if process_id is not None:
            try:
                cleanup_result = _wait_for_owned_process_exit(process_id)
            except Exception as exc:
                cleanup_failures.append(("owned_process_termination", exc))
            else:
                if str(cleanup_result.get("status") or "FAIL") != "PASS":
                    cleanup_failures.append(
                        (
                            "owned_process_termination",
                            RuntimeError(f"owned Excel process cleanup returned {cleanup_result!r}"),
                        )
                    )
        for result in results:
            result.owned_excel_process_cleanup = (
                "FAIL" if cleanup_failures else str(cleanup_result.get("status") or "PASS")
            )
            result.owned_excel_process_forced_termination = bool(
                cleanup_result.get("forced_termination")
            )
            if cleanup_failures:
                _record_cleanup_failures(
                    result,
                    cleanup_failures,
                    started=cleanup_started,
                )
        if cleanup_failures and not results and primary_error is None:
            detail = "; ".join(f"{name}={exc}" for name, exc in cleanup_failures)
            raise RuntimeError(f"Owned Excel visual-render cleanup failed: {detail}")
    return results


def run_render_validation(
    workbooks: Mapping[str, Path | str],
    *,
    output_root: Path | str | None = None,
    timestamp: str | None = None,
    enable_com: bool = True,
    module_manifest_path: Path | str = DEFAULT_MODULE_MANIFEST,
) -> RenderValidationReport:
    started = time.perf_counter()
    ts = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    root = Path(output_root) if output_root is not None else Path(__file__).resolve().parents[2] / "render_checks"
    output_dir = root / f"final_validation_{ts}"
    output_dir.mkdir(parents=True, exist_ok=True)
    normalized = {str(ticker).upper(): Path(path) for ticker, path in workbooks.items()}
    report = RenderValidationReport(output_dir=output_dir, render_status="pending")
    targets_by_ticker = {
        ticker: discover_render_targets(
            workbook_path,
            ticker,
            module_manifest_path=module_manifest_path,
        )
        for ticker, workbook_path in normalized.items()
    }

    for ticker, workbook_path in normalized.items():
        report.style_reports[ticker] = validate_openpyxl_layout(
            workbook_path,
            ticker,
            module_manifest_path=module_manifest_path,
            render_targets=targets_by_ticker[ticker],
        )

    if not enable_com:
        report.render_status = "skipped"
        report.skip_reason = "Excel COM rendering disabled by caller; openpyxl style/layout validation still ran."
    else:
        try:
            report.rendered_ranges = _render_ranges_with_excel_com(
                normalized,
                output_dir,
                targets_by_ticker,
            )
            report.render_status = "pass" if all(result.status == "pass" for result in report.rendered_ranges) else "fail"
        except RenderUnavailableError as exc:
            report.render_status = "skipped"
            report.skip_reason = str(exc)
        except Exception as exc:
            report.render_status = "fail"
            report.skip_reason = f"Excel COM rendering failed: {type(exc).__name__}: {exc}"

    report.elapsed_seconds = time.perf_counter() - started
    _write_reports(report)
    return report


def _write_reports(report: RenderValidationReport) -> None:
    json_path = report.output_dir / "render_validation_report.json"
    csv_path = report.output_dir / "render_validation_summary.csv"
    json_path.write_text(json.dumps(report.to_dict(), indent=2, ensure_ascii=False), encoding="utf-8")
    rows = report.to_summary_rows()
    if rows:
        with csv_path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)


def _default_workbooks(workbook_dir: Path) -> Dict[str, Path]:
    workbooks: Dict[str, Path] = {}
    for ticker in ("PBI", "GPRE", "ANF"):
        workbooks[ticker] = next(
            (
                workbook_dir / f"{ticker}_model{suffix}"
                for suffix in (".xlsm", ".xlsx")
                if (workbook_dir / f"{ticker}_model{suffix}").exists()
            ),
            workbook_dir / f"{ticker}_model.xlsx",
        )
    return workbooks


def resolve_workbook_dir(
    *,
    data_root: Path | str | None = None,
    workbook_dir: Path | str | None = None,
) -> Path:
    if workbook_dir is not None and str(workbook_dir).strip():
        return Path(workbook_dir).expanduser().resolve()
    paths = resolve_stock_model_paths(Path(__file__).resolve().parents[2], data_root)
    return paths.excel_output_dir


def resolve_output_root(
    *,
    data_root: Path | str | None = None,
    output_root: Path | str | None = None,
) -> Path:
    if output_root is not None and str(output_root).strip():
        return Path(output_root).expanduser().resolve()
    paths = resolve_stock_model_paths(Path(__file__).resolve().parents[2], data_root)
    return paths.render_checks_dir


def _main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Render workbook ranges and validate user-facing workbook styles.")
    parser.add_argument("--workbook-dir", default=None)
    parser.add_argument("--output-root", default=None)
    parser.add_argument("--data-root", default="", help="Portable StockModelData root.")
    parser.add_argument("--openpyxl-only", action="store_true", help="Skip Excel COM image rendering but run layout/style validation.")
    args = parser.parse_args(argv)

    workbook_dir = resolve_workbook_dir(data_root=args.data_root, workbook_dir=args.workbook_dir)
    output_root = resolve_output_root(data_root=args.data_root, output_root=args.output_root)
    report = run_render_validation(
        _default_workbooks(workbook_dir),
        output_root=output_root,
        enable_com=not args.openpyxl_only,
    )
    print("Ticker | Style | Render | Overall")
    print("-------+-------+--------+--------")
    for row in report.to_summary_rows():
        print(f"{row['Ticker']:<6} | {row['Style']:<5} | {row['Render']:<6} | {row['Overall']}")
    if report.skip_reason:
        print(f"Render skip reason: {report.skip_reason}")
    print(f"Report: {report.output_dir}")
    return 1 if report.overall == "FAIL" else 0


if __name__ == "__main__":
    raise SystemExit(_main())
