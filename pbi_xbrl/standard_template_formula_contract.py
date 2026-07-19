"""Ticker-neutral formula and visible-label contract for the frozen shell.

Raw, source-backed facts are written by the binding plan.  This module owns only
deterministic workbook formulas and generic presentation labels.  It deliberately
contains no ticker names, source values, or valuation assumptions.
"""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Collection, Mapping

from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Protection
from openpyxl.utils import get_column_letter, quote_sheetname, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

from pbi_xbrl.json_schema_validation import load_json_strict

from pbi_xbrl.valuation_scenario_economics import (
    CANONICAL_DIRECT_REVENUE_PROPAGATION,
    CANONICAL_DIRECT_ROUTE_VALUE_KIND,
    CANONICAL_POPULATED_STATUS,
    CANONICAL_PROFILE_REVENUE_PROPAGATION,
    CANONICAL_PROFILE_ROUTE_VALUE_KIND,
    CANONICAL_REVENUE_OUTPUT_METRIC,
    CANONICAL_REVENUE_UNIT,
    CANONICAL_TOTAL_COMPANY_TOKEN,
)


FORMULA_CONTRACT_VERSION = "2.1.0"
ROOT = Path(__file__).resolve().parents[1]
HIDDEN_VALUE_SIGNAL_CONTRACT = ROOT / "docs" / "hidden_value_signal_contract.json"
HIDDEN_VALUE_DETAIL_FIRST_ROW = 2
HIDDEN_VALUE_CANDIDATE_FIRST_ROW = 2
HIDDEN_VALUE_BASE_LAST_ROW = 5001

INVESTMENT_CASE_SCENARIO_OWNED_RANGES = (
    "A13:K180",
)
FIRST_QUARTER_COLUMN = 2
LAST_QUARTER_COLUMN = 13
CALCULATION_HISTORY_FIRST_ROW = 2
CALCULATION_HISTORY_LAST_ROW = 1000
CALCULATION_HISTORY_COLUMNS = {
    "period": "A",
    "period_ordinal": "B",
    "metric": "C",
    "value": "D",
    "unit": "E",
    "source_ref": "F",
    "status": "G",
}


@dataclass(frozen=True)
class FormulaRow:
    formula_id: str
    sheet: str
    row: int
    number_format: str
    description: str


@dataclass(frozen=True)
class FormulaTargetContract:
    """One formula contract ID and its exact owned workbook cells."""

    formula_id: str
    sheet: str
    targets: tuple[str, ...]


@dataclass(frozen=True)
class UserInputContract:
    """One authoritative editable surface and its Excel validation."""

    input_id: str
    sheet: str
    target: str
    surface_id: str
    surface_target: str
    required_formula_ids: tuple[str, ...]
    validation_type: str
    operator: str | None = None
    formula1: str | None = None
    formula2: str | None = None


USER_INPUT_CONTRACTS = (
    UserInputContract("valuation_price", "Valuation", "D194", "valuation_price", "D194", ("valuation_output_formulas",), "decimal", "greaterThan", "0"),
    UserInputContract("valuation_target_multiples", "Valuation", "D208:D209", "valuation_targets", "D208:D210", ("valuation_output_formulas",), "decimal", "greaterThan", "0"),
    UserInputContract("valuation_target_yield", "Valuation", "D210", "valuation_targets", "D208:D210", ("valuation_output_formulas",), "decimal", "between", "0.000001", "1"),
    UserInputContract("valuation_maintenance_capex_ratio", "Valuation", "D213", "valuation_normalization", "D213:D216", ("valuation_output_formulas",), "decimal", "between", "0", "1"),
    UserInputContract("valuation_recurring_cash_costs", "Valuation", "D214", "valuation_normalization", "D213:D216", ("valuation_output_formulas",), "custom", formula1='=OR(D214="",ISNUMBER(D214))'),
    UserInputContract("valuation_working_capital_normalization", "Valuation", "D215", "valuation_normalization", "D213:D216", ("valuation_output_formulas",), "custom", formula1='=OR(D215="",ISNUMBER(D215))'),
    UserInputContract("valuation_per_share_mode", "Valuation", "D216", "valuation_normalization", "D213:D216", ("valuation_output_formulas",), "list", formula1='"Outstanding,Diluted"'),
    UserInputContract("valuation_scenario_targets", "Valuation", "D218:D219", "valuation_scenario_targets", "D218:D220", ("valuation_scenario_formulas",), "decimal", "greaterThan", "0"),
    UserInputContract("valuation_dcf_horizon", "Valuation", "D220", "valuation_scenario_targets", "D218:D220", ("valuation_scenario_formulas",), "whole", "between", "1", "20"),
    UserInputContract("valuation_dcf_fcff", "Valuation", "J218", "valuation_dcf_inputs", "J218:J221", ("valuation_scenario_formulas",), "decimal", "greaterThan", "0"),
    UserInputContract("valuation_dcf_growth", "Valuation", "J219:J220", "valuation_dcf_inputs", "J218:J221", ("valuation_scenario_formulas",), "decimal", "between", "-1", "1"),
    UserInputContract("valuation_dcf_wacc", "Valuation", "J221", "valuation_dcf_inputs", "J218:J221", ("valuation_scenario_formulas",), "decimal", "between", "0.000001", "1"),
    UserInputContract("valuation_terminal_growth_axis", "Valuation", "H226:L226", "valuation_terminal_growth_axis", "H226:L226", ("valuation_scenario_formulas",), "decimal", "between", "-1", "1"),
    UserInputContract("valuation_wacc_axis", "Valuation", "G227:G234", "valuation_wacc_axis", "G227:G234", ("valuation_scenario_formulas",), "decimal", "between", "0.000001", "1"),
    UserInputContract("valuation_scenario_profile", "Valuation", "E236", "valuation_scenario_assumptions", "E236:E240", ("valuation_scenario_formulas",), "list", formula1='"Bear,Base,Bull,Custom"'),
    UserInputContract("valuation_scenario_growth_margins", "Valuation", "E237:E239", "valuation_scenario_assumptions", "E236:E240", ("valuation_scenario_formulas",), "decimal", "between", "-1", "1"),
    UserInputContract("valuation_scenario_tax_rate", "Valuation", "E240", "valuation_scenario_assumptions", "E236:E240", ("valuation_scenario_formulas",), "decimal", "between", "0", "1"),
    UserInputContract("valuation_scenario_horizon", "Valuation", "J236", "valuation_scenario_horizon", "J236", ("valuation_scenario_formulas",), "custom", formula1='=OR(J236="",AND(LEFT(J236,2)="FY",LEN(J236)=6,ISNUMBER(VALUE(RIGHT(J236,4)))),AND(LEN(J236)=7,MID(J236,5,2)="-Q",ISNUMBER(VALUE(LEFT(J236,4))),VALUE(RIGHT(J236,1))>=1,VALUE(RIGHT(J236,1))<=4))'),
    UserInputContract("valuation_pretax_bridge", "Valuation", "D247", "valuation_tax_bridges", "D247:E248", ("valuation_scenario_formulas",), "custom", formula1='=OR(D247="",ISNUMBER(D247))'),
    UserInputContract("valuation_pretax_bridge_tax", "Valuation", "E247", "valuation_tax_bridges", "D247:E248", ("valuation_scenario_formulas",), "list", formula1='"taxable,non_taxable,non_taxable_credit,cash_only,no_eps_impact"'),
    UserInputContract("valuation_interest_bridge", "Valuation", "D248", "valuation_tax_bridges", "D247:E248", ("valuation_scenario_formulas",), "custom", formula1='=OR(D248="",ISNUMBER(D248))'),
    UserInputContract("valuation_interest_bridge_tax", "Valuation", "E248", "valuation_tax_bridges", "D247:E248", ("valuation_scenario_formulas",), "list", formula1='"taxable,non_taxable"'),
    UserInputContract("valuation_cash_normalization_bridges", "Valuation", "D249:D250", "valuation_cash_normalization", "D249:D250", ("valuation_scenario_formulas",), "custom", formula1='=OR(D249="",ISNUMBER(D249))'),
    UserInputContract("valuation_buyback_cash", "Valuation", "D253", "valuation_capital_allocation", "D253:D256", ("valuation_scenario_formulas",), "decimal", "greaterThanOrEqual", "0"),
    UserInputContract("valuation_buyback_price", "Valuation", "D254", "valuation_capital_allocation", "D253:D256", ("valuation_scenario_formulas",), "decimal", "greaterThan", "0"),
    UserInputContract("valuation_issuance_debt_paydown", "Valuation", "D255:D256", "valuation_capital_allocation", "D253:D256", ("valuation_scenario_formulas",), "decimal", "greaterThanOrEqual", "0"),
    UserInputContract("investment_case_horizon", "{ticker}_Investment_Case", "B23:D23", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "custom", formula1='=OR(B23="",AND(LEFT(B23,2)="FY",LEN(B23)=6,ISNUMBER(VALUE(RIGHT(B23,4)))))'),
    UserInputContract("investment_case_growth_margins", "{ticker}_Investment_Case", "B24:D26", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "decimal", "between", "-1", "1"),
    UserInputContract("investment_case_pretax_bridge", "{ticker}_Investment_Case", "B27:D27", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "custom", formula1='=OR(B27="",ISNUMBER(B27))'),
    UserInputContract("investment_case_tax_rate", "{ticker}_Investment_Case", "B28:D28", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "decimal", "between", "0", "1"),
    UserInputContract("investment_case_bridge_tax", "{ticker}_Investment_Case", "B29:D29", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "list", formula1='"taxable,non_taxable,non_taxable_credit,cash_only,no_eps_impact"'),
    UserInputContract("investment_case_interest_bridge", "{ticker}_Investment_Case", "B30:D30", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "custom", formula1='=OR(B30="",ISNUMBER(B30))'),
    UserInputContract("investment_case_interest_tax", "{ticker}_Investment_Case", "B31:D31", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "list", formula1='"taxable,non_taxable"'),
    UserInputContract("investment_case_cash_bridges", "{ticker}_Investment_Case", "B32:D33", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "custom", formula1='=OR(B32="",ISNUMBER(B32))'),
    UserInputContract("investment_case_buyback_cash", "{ticker}_Investment_Case", "B34:D34", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "decimal", "greaterThanOrEqual", "0"),
    UserInputContract("investment_case_buyback_price", "{ticker}_Investment_Case", "B35:D35", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "decimal", "greaterThan", "0"),
    UserInputContract("investment_case_issuance_paydown", "{ticker}_Investment_Case", "B36:D37", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "decimal", "greaterThanOrEqual", "0"),
    UserInputContract("investment_case_target_multiples", "{ticker}_Investment_Case", "B38:D41", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "decimal", "greaterThan", "0"),
    UserInputContract("investment_case_target_yield", "{ticker}_Investment_Case", "B42:D42", "investment_case_scenarios", "B23:D42", ("investment_case_scenario_formulas",), "decimal", "between", "0.000001", "1"),
    UserInputContract("investment_case_eps_axis", "{ticker}_Investment_Case", "B160:D160", "investment_case_eps_axis", "B160:D160", ("investment_case_sensitivity_formulas",), "decimal", "greaterThan", "0"),
    UserInputContract("investment_case_pe_axis", "{ticker}_Investment_Case", "A161:A163", "investment_case_pe_axis", "A161:A163", ("investment_case_sensitivity_formulas",), "decimal", "greaterThan", "0"),
    UserInputContract("investment_case_adjusted_ebitda_axis", "{ticker}_Investment_Case", "B171:D171", "investment_case_adjusted_ebitda_axis", "B171:D171", ("investment_case_sensitivity_formulas",), "decimal", "greaterThan", "0"),
    UserInputContract("investment_case_ev_multiple_axis", "{ticker}_Investment_Case", "A172:A174", "investment_case_ev_multiple_axis", "A172:A174", ("investment_case_sensitivity_formulas",), "decimal", "greaterThan", "0"),
    UserInputContract("investment_case_fcf_yield_axis", "{ticker}_Investment_Case", "B177:D177", "investment_case_fcf_yield_axis", "B177:D177", ("investment_case_sensitivity_formulas",), "decimal", "between", "0.000001", "1"),
    UserInputContract("investment_case_fcf_axis", "{ticker}_Investment_Case", "A178:A180", "investment_case_fcf_axis", "A178:A180", ("investment_case_sensitivity_formulas",), "decimal", "greaterThan", "0"),
)


VALUATION_RAW_ROWS = {
    "revenue": 9,
    "base_ebitda": 18,
    "adjusted_ebitda": 24,
    "operating_income": 32,
    "net_income": 36,
    "operating_cash_flow": 43,
    "capital_expenditures": 44,
    "interest_paid": 59,
    "buybacks_cash": 62,
    "cash": 70,
    "marketable_securities": 71,
    "debt_core": 72,
    "lease_liabilities": 79,
    "pension_obligation_net": 82,
    "revolver_availability": 95,
    "diluted_shares": 102,
    "shares_outstanding": 103,
    "eps": 107,
    "adjusted_eps": 110,
}

# Hidden rows carry source-backed inputs that have no useful standalone visible
# row.  They are binding targets, not formulas, and remain part of the signed shell.
VALUATION_HELPER_ROWS = {
    "gross_profit": 262,
    "interest_expense": 263,
    "dividends_cash": 264,
    "acquisitions_cash": 265,
    "debt_repayment": 266,
    "debt_issuance": 267,
    "total_equity": 268,
    "goodwill": 269,
    "intangibles": 270,
}

VALUATION_FORMULA_HELPER_ROWS = {
    "operating_cash_flow_ttm": 271,
}

BS_RAW_ROWS = {
    "cash": 9,
    "restricted_cash": 10,
    "marketable_securities": 13,
    "accounts_receivable": 14,
    "inventory": 15,
    "current_assets": 18,
    "property_plant_equipment_net": 19,
    "goodwill": 22,
    "intangibles": 23,
    "other_assets_noncurrent": 24,
    "total_assets": 25,
    "accounts_payable": 28,
    "accrued_liabilities": 29,
    "short_term_borrowings": 32,
    "debt_current": 33,
    "lease_liabilities_current": 34,
    "current_liabilities": 35,
    "debt_core": 40,
    "lease_liabilities_noncurrent": 42,
    "pension_obligation_net": 43,
    "other_liabilities_noncurrent": 44,
    "total_liabilities": 45,
    "total_equity": 47,
    "shares_outstanding": 48,
    "diluted_shares": 49,
}

BS_QUARTERLY_FORMULA_ROWS = (
    FormulaRow("bs_cash_including_restricted", "BS_Segments", 11, "#,##0.0;[Red]-#,##0.0", "Cash plus restricted cash."),
    FormulaRow("bs_cash_qoq", "BS_Segments", 12, "#,##0.0;[Red]-#,##0.0", "Quarter-over-quarter cash change."),
    FormulaRow("bs_goodwill_assets_ratio", "BS_Segments", 26, "0.0%;[Red]-0.0%", "Goodwill divided by total assets."),
    FormulaRow("bs_working_capital", "BS_Segments", 36, "#,##0.0;[Red]-#,##0.0", "Current assets less current liabilities."),
    FormulaRow("bs_working_capital_qoq", "BS_Segments", 37, "#,##0.0;[Red]-#,##0.0", "Quarter-over-quarter working-capital change."),
    FormulaRow("bs_current_ratio", "BS_Segments", 38, "0.0%;[Red]-0.0%", "Current assets divided by current liabilities."),
    FormulaRow("bs_quick_ratio", "BS_Segments", 39, "0.0%;[Red]-0.0%", "Cash-adjusted current assets divided by current liabilities."),
    FormulaRow("bs_debt_qoq", "BS_Segments", 41, "#,##0.0;[Red]-#,##0.0", "Quarter-over-quarter core-debt change."),
    FormulaRow("bs_inventory_yoy", "BS_Segments", 51, "0.0%;[Red]-0.0%", "Inventory growth versus prior year."),
    FormulaRow("bs_revenue_yoy", "BS_Segments", 52, "0.0%;[Red]-0.0%", "Revenue growth versus prior year."),
    FormulaRow("bs_inventory_vs_revenue_growth", "BS_Segments", 53, "0.0%;[Red]-0.0%", "Inventory growth less revenue growth."),
    FormulaRow("bs_core_net_cash", "BS_Segments", 54, "#,##0.0;[Red]-#,##0.0", "Cash plus securities less core debt."),
    FormulaRow("bs_total_lease_liabilities", "BS_Segments", 55, "#,##0.0;[Red]-#,##0.0", "Current plus non-current lease liabilities."),
    FormulaRow("bs_diluted_shares_yoy", "BS_Segments", 56, "0.0%;[Red]-0.0%", "Diluted-share growth versus prior year."),
)

VALUATION_OUTPUT_FORMULA_CELLS = tuple(f"N{row}" for row in range(194, 211))
VALUATION_SIDECAR_FORMULA_CELLS = tuple(f"U{row}" for row in (64, 65, 66, 67, 68, 69, 70, 73, 74, 75))


FORMULA_ROWS = (
    FormulaRow("revenue_ttm", "Valuation", 10, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter revenue."),
    FormulaRow("revenue_yoy", "Valuation", 11, "0.0%;[Red]-0.0%", "Quarterly revenue growth versus prior year."),
    FormulaRow("gross_margin", "Valuation", 13, "0.0%;[Red]-0.0%", "Gross profit divided by revenue."),
    FormulaRow("operating_margin", "Valuation", 14, "0.0%;[Red]-0.0%", "Operating income divided by revenue."),
    FormulaRow("operating_margin_ttm", "Valuation", 15, "0.0%;[Red]-0.0%", "TTM operating income divided by TTM revenue."),
    FormulaRow("ebitda_margin", "Valuation", 19, "0.0%;[Red]-0.0%", "Base EBITDA divided by revenue."),
    FormulaRow("ebitda_yoy", "Valuation", 20, "0.0%;[Red]-0.0%", "Base EBITDA growth versus prior year."),
    FormulaRow("ebitda_ttm", "Valuation", 21, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter base EBITDA."),
    FormulaRow("ebitda_margin_ttm", "Valuation", 22, "0.0%;[Red]-0.0%", "TTM base EBITDA divided by TTM revenue."),
    FormulaRow("adjusted_ebitda_ttm", "Valuation", 25, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter adjusted EBITDA."),
    FormulaRow("adjusted_ebitda_delta", "Valuation", 26, "#,##0.0;[Red]-#,##0.0", "Adjusted EBITDA less base EBITDA."),
    FormulaRow("adjusted_ebitda_margin", "Valuation", 27, "0.0%;[Red]-0.0%", "Adjusted EBITDA divided by revenue."),
    FormulaRow("adjusted_ebitda_yoy", "Valuation", 28, "0.0%;[Red]-0.0%", "Adjusted EBITDA growth versus prior year."),
    FormulaRow("adjusted_ebitda_margin_ttm", "Valuation", 29, "0.0%;[Red]-0.0%", "TTM adjusted EBITDA divided by TTM revenue."),
    FormulaRow("operating_income_margin", "Valuation", 33, "0.0%;[Red]-0.0%", "Operating income divided by revenue."),
    FormulaRow("operating_income_ttm", "Valuation", 34, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter operating income."),
    FormulaRow("operating_income_margin_ttm", "Valuation", 35, "0.0%;[Red]-0.0%", "TTM operating income divided by TTM revenue."),
    FormulaRow("net_margin", "Valuation", 37, "0.0%;[Red]-0.0%", "Net income divided by revenue."),
    FormulaRow("net_income_yoy", "Valuation", 38, "0.0%;[Red]-0.0%", "Net income growth versus prior year."),
    FormulaRow("net_income_ttm", "Valuation", 39, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter net income."),
    FormulaRow("net_margin_ttm", "Valuation", 40, "0.0%;[Red]-0.0%", "TTM net income divided by TTM revenue."),
    FormulaRow("capex_margin", "Valuation", 45, "0.0%;[Red]-0.0%", "Capital expenditures divided by revenue."),
    FormulaRow("capex_margin_ttm", "Valuation", 46, "0.0%;[Red]-0.0%", "TTM capital expenditures divided by TTM revenue."),
    FormulaRow("free_cash_flow", "Valuation", 47, "#,##0.0;[Red]-#,##0.0", "Operating cash flow less capital expenditures."),
    FormulaRow("free_cash_flow_yoy_delta", "Valuation", 48, "#,##0.0;[Red]-#,##0.0", "Free-cash-flow change versus prior year."),
    FormulaRow("free_cash_flow_ttm", "Valuation", 49, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter free cash flow."),
    FormulaRow("free_cash_flow_ttm_yoy_delta", "Valuation", 50, "#,##0.0;[Red]-#,##0.0", "TTM free-cash-flow change versus prior year."),
    FormulaRow("free_cash_flow_margin", "Valuation", 57, "0.0%;[Red]-0.0%", "Free cash flow divided by revenue."),
    FormulaRow("free_cash_flow_margin_ttm", "Valuation", 58, "0.0%;[Red]-0.0%", "TTM free cash flow divided by TTM revenue."),
    FormulaRow("buybacks_ttm", "Valuation", 63, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter cash used for buybacks."),
    FormulaRow("dividends_ttm", "Valuation", 64, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter cash dividends."),
    FormulaRow("acquisitions_ttm", "Valuation", 65, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter acquisition cash outflow."),
    FormulaRow("debt_repayment_ttm", "Valuation", 66, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter gross debt repayment."),
    FormulaRow("debt_issuance_ttm", "Valuation", 67, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter gross debt issuance."),
    FormulaRow("net_debt", "Valuation", 73, "#,##0.0;[Red]-#,##0.0", "Core debt less cash."),
    FormulaRow("net_debt_qoq", "Valuation", 74, "#,##0.0;[Red]-#,##0.0", "Quarter-over-quarter change in net debt."),
    FormulaRow("net_debt_yoy", "Valuation", 75, "#,##0.0;[Red]-#,##0.0", "Year-over-year change in net debt."),
    FormulaRow("core_net_cash", "Valuation", 77, "#,##0.0;[Red]-#,##0.0", "Cash less core debt."),
    FormulaRow("net_cash_with_securities", "Valuation", 78, "#,##0.0;[Red]-#,##0.0", "Cash plus marketable securities less core debt."),
    FormulaRow("lease_adjusted_net_debt", "Valuation", 80, "#,##0.0;[Red]-#,##0.0", "Core debt plus lease liabilities less cash."),
    FormulaRow("lease_adjusted_net_debt_with_securities", "Valuation", 81, "#,##0.0;[Red]-#,##0.0", "Lease-adjusted net debt after marketable securities."),
    FormulaRow("ebitda_ttm_copy", "Valuation", 84, "#,##0.0;[Red]-#,##0.0", "Coverage-panel base EBITDA TTM."),
    FormulaRow("adjusted_ebitda_ttm_copy", "Valuation", 85, "#,##0.0;[Red]-#,##0.0", "Coverage-panel adjusted EBITDA TTM."),
    FormulaRow("net_leverage", "Valuation", 86, "0.00x;[Red]-0.00x", "Net debt divided by base EBITDA TTM."),
    FormulaRow("adjusted_net_leverage", "Valuation", 87, "0.00x;[Red]-0.00x", "Net debt divided by adjusted EBITDA TTM."),
    FormulaRow("interest_coverage", "Valuation", 88, "0.00x;[Red]-0.00x", "TTM operating income divided by TTM interest expense."),
    FormulaRow("cash_interest_coverage", "Valuation", 89, "0.00x;[Red]-0.00x", "TTM base EBITDA divided by TTM cash interest paid."),
    FormulaRow("fcf_conversion", "Valuation", 90, "0.0%;[Red]-0.0%", "TTM free cash flow divided by TTM base EBITDA."),
    FormulaRow("diluted_shares_qoq", "Valuation", 104, "0.000;[Red]-0.000", "Quarter-over-quarter diluted share change."),
    FormulaRow("diluted_shares_yoy", "Valuation", 105, "0.000;[Red]-0.000", "Year-over-year diluted share change."),
    FormulaRow("gaap_eps_yoy", "Valuation", 108, "0.0%;[Red]-0.0%", "GAAP diluted EPS growth versus prior year."),
    FormulaRow("gaap_eps_ttm", "Valuation", 109, "$0.00;[Red]-$0.00", "Trailing-four-quarter GAAP diluted EPS."),
    FormulaRow("adjusted_eps_ttm", "Valuation", 111, "$0.00;[Red]-$0.00", "Trailing-four-quarter adjusted diluted EPS."),
    FormulaRow("book_value_per_share", "Valuation", 113, "$0.00;[Red]-$0.00", "Total equity divided by point-in-time shares outstanding."),
    FormulaRow("tangible_book_value_per_share", "Valuation", 114, "$0.00;[Red]-$0.00", "Equity less goodwill and intangibles, divided by point-in-time shares outstanding."),
    FormulaRow("free_cash_flow_per_share", "Valuation", 115, "$0.00;[Red]-$0.00", "TTM free cash flow divided by diluted shares."),
    FormulaRow("operating_cash_flow_ttm", "Valuation", 271, "#,##0.0;[Red]-#,##0.0", "Trailing-four-quarter operating cash flow."),
)


def formula_target_contracts() -> tuple[FormulaTargetContract, ...]:
    """Return the complete executable formula-cell contract for the union shell."""

    contracts = [
        FormulaTargetContract(row.formula_id, row.sheet, (f"B{row.row}:M{row.row}",))
        for row in FORMULA_ROWS
    ]
    contracts.extend(
        FormulaTargetContract(row.formula_id, row.sheet, (f"B{row.row}:M{row.row}",))
        for row in BS_QUARTERLY_FORMULA_ROWS
    )
    contracts.extend(
        (
            FormulaTargetContract("summary_revenue_ttm", "SUMMARY", ("B27",)),
            FormulaTargetContract("summary_revenue_yoy", "SUMMARY", ("B29",)),
            FormulaTargetContract("summary_net_income_yoy", "SUMMARY", ("B31",)),
            FormulaTargetContract("summary_gaap_eps", "SUMMARY", ("B32",)),
            FormulaTargetContract("summary_gaap_eps_yoy", "SUMMARY", ("B33",)),
            FormulaTargetContract("summary_free_cash_flow_ttm", "SUMMARY", ("B36",)),
            FormulaTargetContract("summary_free_cash_flow_yoy", "SUMMARY", ("B37",)),
            FormulaTargetContract("summary_interest_coverage", "SUMMARY", ("B42",)),
            FormulaTargetContract("valuation_output_formulas", "Valuation", ("N194:N210",)),
            FormulaTargetContract("valuation_sidecar_formulas", "Valuation", ("U64:U70", "U72:U75")),
            FormulaTargetContract(
                "valuation_scenario_formulas",
                "Valuation",
                (
                    "Q219:Q221",
                    "J222:J223",
                    "H227:L234",
                    "F236",
                    "E241:E244",
                    "E259:E261",
                    "J259:J261",
                    "N236:N244",
                    "N259:N261",
                    "R259:R261",
                ),
            ),
            FormulaTargetContract(
                "scenario_revenue_route_formulas",
                "Valuation_Summary",
                ("H2:K2",),
            ),
            FormulaTargetContract(
                "investment_case_scenario_formulas",
                "{ticker}_Investment_Case",
                ("B15:D22", "B50:D53", "B56:D58", "B62:B68", "B85:J87"),
            ),
            FormulaTargetContract(
                "investment_case_sensitivity_formulas",
                "{ticker}_Investment_Case",
                ("B161:D163", "B172:D174", "B178:D180"),
            ),
            FormulaTargetContract(
                "valuation_summary_formulas",
                "Valuation_Summary",
                ("B2:B20", "D2:D20", "F2:F20"),
            ),
            FormulaTargetContract(
                "valuation_grid_formulas",
                "Valuation_Grid",
                ("A2:A42", "B2:B42", "C2:C42", "E2:F42"),
            ),
            FormulaTargetContract(
                "hidden_value_recompute_detail_formulas",
                "Hidden_Value_Recompute",
                ("X2:AD92",),
            ),
            FormulaTargetContract(
                "hidden_value_recompute_candidate_formulas",
                "Hidden_Value_Recompute",
                ("AF2:AX8",),
            ),
            FormulaTargetContract(
                "hidden_value_audit_parity_formulas",
                "Hidden_Value_Audit",
                ("R2:V8",),
            ),
            FormulaTargetContract(
                "hidden_value_valuation_state_count_formulas",
                "Valuation",
                ("R139:R143",),
            ),
        )
    )
    return tuple(contracts)


def active_user_input_contracts(enabled_formula_ids: Collection[str]) -> tuple[UserInputContract, ...]:
    enabled = {str(value) for value in enabled_formula_ids}
    return tuple(
        contract
        for contract in USER_INPUT_CONTRACTS
        if set(contract.required_formula_ids) <= enabled
    )


def user_input_surface_targets(sheet: str) -> tuple[str, ...]:
    """Return broad shell ownership ranges projected from the input authority."""

    return tuple(
        dict.fromkeys(
            contract.surface_target
            for contract in USER_INPUT_CONTRACTS
            if contract.sheet == sheet
        )
    )


def _resolved_contract_sheet(workbook: Any, sheet_name: str) -> str | None:
    if sheet_name in workbook.sheetnames:
        return sheet_name
    if "{ticker}" not in sheet_name:
        return None
    suffix = sheet_name.replace("{ticker}", "")
    candidates = [name for name in workbook.sheetnames if name.endswith(suffix) and "{ticker}" not in name]
    return candidates[0] if len(candidates) == 1 else None


def _cells_for_target(ws: Any, target: str) -> tuple[Any, ...]:
    min_col, min_row, max_col, max_row = range_boundaries(target)
    return tuple(
        cell
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
        for cell in row
        if not isinstance(cell, MergedCell)
    )


def apply_workbook_protection_contract(workbook: Any, enabled_formula_ids: Collection[str]) -> None:
    """Lock the workbook and expose only declared, validated user inputs."""

    active = active_user_input_contracts(enabled_formula_ids)
    contract_sheets = {contract.sheet for contract in USER_INPUT_CONTRACTS}
    for ws in workbook.worksheets:
        for cell in tuple(ws._cells.values()):
            protection = copy(cell.protection)
            protection.locked = True
            cell.protection = protection
        ws.protection.sheet = True
        ws.protection.objects = True
        ws.protection.scenarios = True
        ws.protection.selectLockedCells = False
        ws.protection.selectUnlockedCells = False

    for sheet_name in contract_sheets:
        resolved = _resolved_contract_sheet(workbook, sheet_name)
        if resolved is None:
            continue
        ws = workbook[resolved]
        _remove_data_validations_overlapping(
            ws,
            tuple(contract.target for contract in USER_INPUT_CONTRACTS if contract.sheet == sheet_name),
        )

    for contract in active:
        resolved = _resolved_contract_sheet(workbook, contract.sheet)
        if resolved is None:
            raise ValueError(f"Active user-input contract requires missing sheet {contract.sheet!r}.")
        ws = workbook[resolved]
        for cell in _cells_for_target(ws, contract.target):
            protection = copy(cell.protection)
            protection.locked = False
            cell.protection = protection
        _add_validation(
            ws,
            contract.validation_type,
            contract.target,
            operator=contract.operator,
            formula1=contract.formula1,
            formula2=contract.formula2,
        )


def workbook_protection_payload(workbook: Any) -> dict[str, Any]:
    worksheets = []
    editable_cells = []
    for ws in workbook.worksheets:
        protection = ws.protection
        worksheets.append(
            {
                "sheet": ws.title,
                "state": ws.sheet_state,
                "sheet_protected": bool(protection.sheet),
                "objects": bool(protection.objects),
                "scenarios": bool(protection.scenarios),
                "select_locked_cells": bool(protection.selectLockedCells),
                "select_unlocked_cells": bool(protection.selectUnlockedCells),
            }
        )
        editable_cells.extend(
            {"sheet": ws.title, "cell": cell.coordinate}
            for cell in tuple(ws._cells.values())
            if not bool(True if cell.protection.locked is None else cell.protection.locked)
        )
    return {
        "worksheets": worksheets,
        "editable_cells": sorted(editable_cells, key=lambda row: (row["sheet"], row["cell"])),
    }


def _validation_formula_at_cell(value: Any, *, origin: str, target: str) -> str:
    raw = str(value or "")
    if not raw:
        return ""
    expression = raw if raw.startswith("=") else f"={raw}"
    try:
        translated = Translator(expression, origin=origin).translate_formula(target)
    except Exception:
        translated = expression
    return translated[1:] if translated.startswith("=") else translated


def _validation_operator(validation_type: str, operator: Any, formula2: Any) -> str:
    value = str(operator or "")
    if not value and validation_type in {"decimal", "whole", "date", "time", "textLength"} and formula2 not in (None, ""):
        return "between"
    return value


def canonical_data_validation_cells(worksheet: Any) -> list[dict[str, str]]:
    """Expand Excel-grouped validations to one semantic contract per cell."""

    result: list[dict[str, str]] = []
    for validation in worksheet.data_validations.dataValidation:
        serialized_ranges = str(validation.sqref).split()
        if not serialized_ranges:
            continue
        first_min_col, first_min_row, _first_max_col, _first_max_row = range_boundaries(serialized_ranges[0])
        origin = f"{get_column_letter(first_min_col)}{first_min_row}"
        validation_type = str(validation.type or "")
        operator = _validation_operator(validation_type, validation.operator, validation.formula2)
        for target_range in serialized_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(target_range)
            for row in range(min_row, max_row + 1):
                for column in range(min_col, max_col + 1):
                    coordinate = f"{get_column_letter(column)}{row}"
                    result.append(
                        {
                            "cell": coordinate,
                            "type": validation_type,
                            "operator": operator,
                            "formula1": _validation_formula_at_cell(
                                validation.formula1,
                                origin=origin,
                                target=coordinate,
                            ),
                            "formula2": _validation_formula_at_cell(
                                validation.formula2,
                                origin=origin,
                                target=coordinate,
                            ),
                        }
                    )
    return sorted(
        result,
        key=lambda row: (row["cell"], row["type"], row["operator"], row["formula1"], row["formula2"]),
    )


def validate_workbook_protection_contract(
    workbook: Any,
    enabled_formula_ids: Collection[str],
) -> list[dict[str, str]]:
    """Validate exact editability, validation, visibility and sheet protection."""

    active = active_user_input_contracts(enabled_formula_ids)
    expected_editable: set[tuple[str, str]] = set()
    issues: list[dict[str, str]] = []
    validations_by_sheet = {
        ws.title: canonical_data_validation_cells(ws)
        for ws in workbook.worksheets
    }
    for ws in workbook.worksheets:
        if not ws.protection.sheet:
            issues.append(
                {
                    "rule_id": "worksheet_protection_missing",
                    "message": f"Worksheet {ws.title!r} is not protected.",
                    "sheet": ws.title,
                    "target": "",
                }
            )
    for contract in active:
        resolved = _resolved_contract_sheet(workbook, contract.sheet)
        if resolved is None:
            issues.append(
                {
                    "rule_id": "user_input_sheet_missing",
                    "message": f"Active user input {contract.input_id!r} requires missing sheet {contract.sheet!r}.",
                    "sheet": contract.sheet,
                    "target": contract.target,
                }
            )
            continue
        ws = workbook[resolved]
        contract_cells = tuple(_cells_for_target(ws, contract.target))
        expected_editable.update((resolved, cell.coordinate) for cell in contract_cells)
        min_col, min_row, _max_col, _max_row = range_boundaries(contract.target)
        origin = f"{get_column_letter(min_col)}{min_row}"
        actual_validations = validations_by_sheet[resolved]
        missing_cells: list[str] = []
        duplicate_cells: list[str] = []
        for cell in contract_cells:
            expected_validation = {
                "cell": cell.coordinate,
                "type": contract.validation_type,
                "operator": _validation_operator(contract.validation_type, contract.operator, contract.formula2),
                "formula1": _validation_formula_at_cell(
                    contract.formula1,
                    origin=origin,
                    target=cell.coordinate,
                ),
                "formula2": _validation_formula_at_cell(
                    contract.formula2,
                    origin=origin,
                    target=cell.coordinate,
                ),
            }
            matches = [row for row in actual_validations if row == expected_validation]
            if not matches:
                missing_cells.append(cell.coordinate)
            elif len(matches) > 1:
                duplicate_cells.append(cell.coordinate)
        if missing_cells:
            issues.append(
                {
                    "rule_id": "user_input_validation_missing",
                    "message": (
                        f"User input {contract.input_id!r} lacks its exact validation contract at "
                        f"{', '.join(missing_cells[:8])}."
                    ),
                    "sheet": resolved,
                    "target": contract.target,
                }
            )
        if duplicate_cells:
            issues.append(
                {
                    "rule_id": "user_input_validation_duplicate",
                    "message": (
                        f"User input {contract.input_id!r} has duplicate validation coverage at "
                        f"{', '.join(duplicate_cells[:8])}."
                    ),
                    "sheet": resolved,
                    "target": contract.target,
                }
            )

    actual_editable = {
        (ws.title, cell.coordinate)
        for ws in workbook.worksheets
        for cell in tuple(ws._cells.values())
        if not bool(True if cell.protection.locked is None else cell.protection.locked)
    }
    for sheet, cell in sorted(actual_editable - expected_editable):
        issues.append(
            {
                "rule_id": "unexpected_editable_cell",
                "message": "Cell is editable but has no active UserInputContract.",
                "sheet": sheet,
                "target": cell,
            }
        )
    for sheet, cell in sorted(expected_editable - actual_editable):
        issues.append(
            {
                "rule_id": "declared_user_input_locked",
                "message": "Declared user input is locked.",
                "sheet": sheet,
                "target": cell,
            }
        )
    return issues


def apply_standard_formula_contracts(
    workbook: Any,
    *,
    enabled_formula_ids: Collection[str] | None = None,
) -> None:
    """Apply generic labels, formulas, helper rows, and formula protection."""

    all_formula_ids = {contract.formula_id for contract in formula_target_contracts()}
    enabled = all_formula_ids if enabled_formula_ids is None else {str(value) for value in enabled_formula_ids}
    unknown = sorted(enabled - all_formula_ids)
    if unknown:
        raise ValueError(f"Unknown standard-template formula contracts: {unknown!r}.")

    valuation = workbook["Valuation"]
    bs = workbook["BS_Segments"]
    _clear_disabled_formula_targets(workbook, enabled)
    _prepare_calculation_history_sheet(workbook)
    _extend_balance_sheet_quarterly_axis(bs)
    _prepare_raw_targets(valuation, bs)
    _apply_visible_labels(valuation)
    _apply_hidden_value_visible_layout(valuation)
    _apply_balance_sheet_labels(bs)
    _apply_hidden_helpers(valuation)
    _apply_hidden_formula_helpers(valuation)
    _apply_valuation_quarterly_formulas(valuation, enabled)
    if "SUMMARY" in workbook.sheetnames:
        _apply_summary_formulas(workbook["SUMMARY"], enabled)
    _apply_balance_sheet_formulas(bs, enabled)
    _apply_valuation_input_outputs(valuation, enabled)
    _apply_valuation_sidecar_outputs(valuation, enabled)
    if "{ticker}_Investment_Case" in workbook.sheetnames:
        _apply_investment_case_scenario_formulas(workbook["{ticker}_Investment_Case"], enabled)
    if "Valuation_Summary" in workbook.sheetnames:
        _apply_valuation_summary_formulas(workbook["Valuation_Summary"], enabled)
    if "Valuation_Grid" in workbook.sheetnames:
        _apply_valuation_grid_formulas(workbook["Valuation_Grid"], enabled)
    _apply_scenario_defined_names(workbook, enabled)


def apply_standard_support_formula_contracts(
    workbook: Any,
    *,
    enabled_formula_ids: Collection[str] | None = None,
) -> None:
    """Apply formula-owned projections after hidden support sheets are rebuilt."""

    all_formula_ids = {contract.formula_id for contract in formula_target_contracts()}
    enabled = all_formula_ids if enabled_formula_ids is None else {str(value) for value in enabled_formula_ids}
    unknown = sorted(enabled - all_formula_ids)
    if unknown:
        raise ValueError(f"Unknown standard-template formula contracts: {unknown!r}.")
    if "Valuation_Summary" in workbook.sheetnames:
        _apply_valuation_summary_formulas(workbook["Valuation_Summary"], enabled)
    if "Valuation_Grid" in workbook.sheetnames:
        _apply_valuation_grid_formulas(workbook["Valuation_Grid"], enabled)
    _apply_scenario_revenue_route_formulas(workbook, enabled)
    _apply_hidden_value_support_formulas(workbook, enabled)


def _apply_scenario_revenue_route_formulas(workbook: Any, enabled_formula_ids: set[str]) -> None:
    if "Scenario_Driver_Assumptions" in workbook.sheetnames:
        for coordinate in ("R2", "S2", "T2", "U2"):
            workbook["Scenario_Driver_Assumptions"][coordinate].value = None
    if "Valuation_Summary" not in workbook.sheetnames:
        return
    ws = workbook["Valuation_Summary"]
    targets = {
        "H2": ("bear", "'{ticker}_Investment_Case'!$B$24", "'{ticker}_Investment_Case'!$B$23"),
        "I2": ("base", "'{ticker}_Investment_Case'!$C$24", "'{ticker}_Investment_Case'!$C$23"),
        "J2": ("bull", "'{ticker}_Investment_Case'!$D$24", "'{ticker}_Investment_Case'!$D$23"),
        "K2": ("custom", "ScenarioGrowth", "ScenarioHorizon"),
    }
    if "scenario_revenue_route_formulas" not in enabled_formula_ids:
        for coordinate in targets:
            ws[coordinate].value = None
        return
    for coordinate, (scenario_id, user_growth, scenario_horizon) in targets.items():
        _set_formula(
            ws[coordinate],
            _scenario_revenue_route_formula(
                scenario_id=scenario_id,
                user_growth=user_growth,
                scenario_horizon=scenario_horizon,
            ),
            "0.0%;[Red]-0.0%",
        )


def _scenario_revenue_route_formula(*, scenario_id: str, user_growth: str, scenario_horizon: str) -> str:
    assumptions = quote_sheetname("Scenario_Driver_Assumptions")
    bridges = quote_sheetname("{ticker}_Investment_Case_Data")
    item = {column: f"{assumptions}!${column}$2:${column}$201" for column in "ABCDEFGHIJKLMNOPQ"}
    bridge = {column: f"{bridges}!${column}$2:${column}$201" for column in "ABCDEFGHIJKLMNOPQR"}
    item_dimension = f'--({item["J"]}="{CANONICAL_TOTAL_COMPANY_TOKEN}")'
    item_member = f'--({item["K"]}="{CANONICAL_TOTAL_COMPANY_TOKEN}")'
    bridge_dimension = f'--({bridge["J"]}="{CANONICAL_TOTAL_COMPANY_TOKEN}")'
    bridge_member = f'--({bridge["K"]}="{CANONICAL_TOTAL_COMPANY_TOKEN}")'
    return (
        "=LET("
        f'scenarioKey,"{scenario_id}",userGrowth,{user_growth},scenarioHorizon,{scenario_horizon},'
        f"itemDimension,{item_dimension},itemMember,{item_member},"
        f'directMask,--({item["A"]}=scenarioKey)*--({item["B"]}="{CANONICAL_REVENUE_OUTPUT_METRIC}")*--({item["D"]}="{CANONICAL_DIRECT_ROUTE_VALUE_KIND}")*--({item["H"]}="{CANONICAL_REVENUE_UNIT}")*--({item["I"]}=scenarioHorizon)*itemDimension*itemMember*--({item["M"]}="{CANONICAL_DIRECT_REVENUE_PROPAGATION}")*--({item["N"]}="{CANONICAL_POPULATED_STATUS}")*--({item["Q"]}=""),'
        f'profileMask,--({item["A"]}=scenarioKey)*--({item["B"]}="{CANONICAL_REVENUE_OUTPUT_METRIC}")*--({item["D"]}="{CANONICAL_PROFILE_ROUTE_VALUE_KIND}")*--({item["E"]}="")*--({item["H"]}="{CANONICAL_REVENUE_UNIT}")*--({item["I"]}=scenarioHorizon)*itemDimension*itemMember*--({item["M"]}="{CANONICAL_PROFILE_REVENUE_PROPAGATION}")*--({item["N"]}="{CANONICAL_POPULATED_STATUS}")*--({item["Q"]}<>""),'
        "directCount,SUMPRODUCT(directMask),profileCount,SUMPRODUCT(profileMask),"
        f'directValue,SUMPRODUCT(directMask,{item["E"]}),profilePack,IF(profileCount=1,IFERROR(LOOKUP(2,1/profileMask,{item["Q"]}),""),""),'
        f"bridgeDimension,{bridge_dimension},bridgeMember,{bridge_member},"
        f'bridgeImpact,--({bridge["F"]}="{CANONICAL_REVENUE_OUTPUT_METRIC}"),'
        f'bridgeMask,--({bridge["A"]}=scenarioKey)*--({bridge["C"]}=profilePack)*bridgeImpact*--({bridge["H"]}="{CANONICAL_REVENUE_UNIT}")*--({bridge["I"]}=scenarioHorizon)*bridgeDimension*bridgeMember*--({bridge["O"]}="{CANONICAL_PROFILE_REVENUE_PROPAGATION}")*--({bridge["P"]}="{CANONICAL_POPULATED_STATUS}"),'
        f'bridgeCount,SUMPRODUCT(bridgeMask),bridgeValue,SUMPRODUCT(bridgeMask,{bridge["G"]}),userCount,--(userGrowth<>""),'
        'IF(OR(scenarioHorizon="",directCount+profileCount+userCount<>1),"",'
        'IF(directCount=1,directValue,IF(profileCount=1,IF(bridgeCount=1,bridgeValue,""),IF(ISNUMBER(userGrowth),userGrowth,"")))))'
    )
def _clear_disabled_formula_targets(workbook: Any, enabled_formula_ids: set[str]) -> None:
    from openpyxl.utils.cell import range_boundaries

    for contract in formula_target_contracts():
        if contract.formula_id in enabled_formula_ids or contract.sheet not in workbook.sheetnames:
            continue
        ws = workbook[contract.sheet]
        for target in contract.targets:
                min_col, min_row, max_col, max_row = range_boundaries(target)
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        if isinstance(cell, MergedCell):
                            continue
                        cell.value = None


def _prepare_raw_targets(valuation: Any, bs: Any) -> None:
    for row in sorted(set(VALUATION_RAW_ROWS.values()) | set(VALUATION_HELPER_ROWS.values())):
        for column in range(FIRST_QUARTER_COLUMN, LAST_QUARTER_COLUMN + 1):
            cell = valuation.cell(row, column)
            cell.value = None
            cell.protection = Protection(locked=True)
    for column in range(FIRST_QUARTER_COLUMN, LAST_QUARTER_COLUMN + 1):
        valuation.cell(6, column).protection = Protection(locked=True)
    for row in BS_RAW_ROWS.values():
        for column in range(FIRST_QUARTER_COLUMN, LAST_QUARTER_COLUMN + 1):
            cell = bs.cell(row, column)
            cell.value = None
            cell.protection = Protection(locked=True)
    for column in range(FIRST_QUARTER_COLUMN, LAST_QUARTER_COLUMN + 1):
        bs.cell(7, column).protection = Protection(locked=True)


def _prepare_calculation_history_sheet(workbook: Any) -> None:
    ws = workbook["History_Q"] if "History_Q" in workbook.sheetnames else workbook.create_sheet("History_Q")
    headers = ["period", "period_ordinal", "metric", "value", "unit", "source_ref", "status"]
    for column, header in enumerate(headers, start=1):
        ws.cell(1, column).value = header
        ws.cell(1, column).protection = Protection(locked=True)
    for row in range(CALCULATION_HISTORY_FIRST_ROW, CALCULATION_HISTORY_LAST_ROW + 1):
        for column in range(1, len(headers) + 1):
            cell = ws.cell(row, column)
            cell.value = None
            cell.protection = Protection(locked=True)
    ws.sheet_state = "hidden"


def _apply_visible_labels(ws: Any) -> None:
    labels = {
        18: "EBITDA (base)",
        19: "EBITDA margin %",
        20: "EBITDA YoY %",
        21: "EBITDA (base, TTM)",
        22: "EBITDA margin (TTM)",
        24: "Adjusted EBITDA",
        25: "Adjusted EBITDA (TTM)",
        26: "Adjusted EBITDA - EBITDA",
        27: "Adjusted EBITDA margin %",
        28: "Adjusted EBITDA YoY %",
        29: "Adjusted EBITDA margin (TTM)",
        36: "Net income",
        37: "Net margin %",
        38: "Net income YoY %",
        39: "Net income (TTM)",
        40: "Net margin (TTM)",
        44: "Capex",
        47: "FCF (CFO - Capex)",
        102: "Diluted shares (m)",
        103: "Shares outstanding (m)",
        107: "Diluted EPS (GAAP)",
        108: "Diluted EPS YoY %",
        109: "Diluted EPS (GAAP, TTM)",
        110: "Adjusted diluted EPS",
        111: "Adjusted diluted EPS (TTM)",
        113: "Book value/share",
        114: "Tangible book value/share",
        115: "FCF/share (TTM)",
    }
    for row, label in labels.items():
        ws.cell(row, 1).value = label


def _apply_hidden_value_visible_layout(ws: Any) -> None:
    """Install the neutral five-row Hidden Value display and typed count panel."""

    headers = {
        "A137": "Hidden value flags",
        "A138": "Triggered signal",
        "B138": "Summary",
        "F138": "Score",
        "G138": "State",
        "H138": "As of period",
        "N137": "Signal state counts",
        "N138": "State",
        "R138": "Count",
    }
    for coordinate, value in headers.items():
        ws[coordinate].value = value

    for row in range(139, 144):
        for column in (1, 2, 6, 7, 8):
            ws.cell(row, column).value = None
        ws.row_dimensions[row].height = 28.0
        ws.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row, 6).alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        ws.cell(row, 7).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        ws.cell(row, 8).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for row, label in enumerate(
        (
            "Triggered",
            "Near miss",
            "Not triggered",
            "Insufficient evidence",
            "Unavailable / invalid",
        ),
        start=139,
    ):
        ws.cell(row, 14).value = label
        ws.cell(row, 18).value = None


def _apply_balance_sheet_labels(ws: Any) -> None:
    labels = {
        13: "Short-term investments / marketable securities",
        32: "Short-term borrowings",
        33: "Current maturities of long-term debt",
        40: "Debt (core borrowings)",
    }
    for row, label in labels.items():
        ws.cell(row, 1).value = label


def _extend_balance_sheet_quarterly_axis(ws: Any) -> None:
    """Extend the reusable quarterly BS/segment axis to twelve periods."""

    merge_updates = {
        "A4:I4": "A4:M4",
        "A5:I5": "A5:M5",
        "B6:I6": "B6:M6",
        "A59:I59": "A59:M59",
    }
    existing = {str(item) for item in ws.merged_cells.ranges}
    for old, new in merge_updates.items():
        if old in existing:
            ws.unmerge_cells(old)
        if new not in {str(item) for item in ws.merged_cells.ranges}:
            ws.merge_cells(new)

    source_width = ws.column_dimensions["I"].width
    for column in range(10, LAST_QUARTER_COLUMN + 1):
        letter = get_column_letter(column)
        ws.column_dimensions[letter].width = source_width
        for row in range(4, 69):
            source = ws.cell(row, 9)
            target = ws.cell(row, column)
            target._style = copy(source._style)
            target.number_format = source.number_format
            target.protection = copy(source.protection)


def _apply_hidden_helpers(ws: Any) -> None:
    for metric, row in VALUATION_HELPER_ROWS.items():
        ws.cell(row, 1).value = f"Formula input: {metric}"
        ws.row_dimensions[row].hidden = True
        for column in range(FIRST_QUARTER_COLUMN, LAST_QUARTER_COLUMN + 1):
            cell = ws.cell(row, column)
            cell.value = None
            cell.number_format = "#,##0.0;[Red]-#,##0.0"
            cell.protection = Protection(locked=True)


def _apply_hidden_formula_helpers(ws: Any) -> None:
    for metric, row in VALUATION_FORMULA_HELPER_ROWS.items():
        ws.cell(row, 1).value = f"Formula output: {metric}"
        ws.row_dimensions[row].hidden = True
        for column in range(FIRST_QUARTER_COLUMN, LAST_QUARTER_COLUMN + 1):
            cell = ws.cell(row, column)
            cell.value = None
            cell.number_format = "#,##0.0;[Red]-#,##0.0"
            cell.protection = Protection(locked=True)


def _summary_valuation_value_formula(source_row: int) -> str:
    headers = "'Valuation'!$B$6:$M$6"
    values = f"'Valuation'!$B${source_row}:$M${source_row}"
    match = f"MATCH($B$26,{headers},0)"
    resolved = f"INDEX({values},1,{match})"
    return (
        f'=IFERROR(IF(OR($B$26="",COUNTIF({headers},$B$26)<>1,'
        f'NOT(ISNUMBER({resolved}))),"",{resolved}),"")'
    )


def _summary_free_cash_flow_yoy_formula() -> str:
    headers = "'Valuation'!$B$6:$M$6"
    delta_values = "'Valuation'!$B$48:$M$48"
    fcf_values = "'Valuation'!$B$47:$M$47"
    match = f"MATCH($B$26,{headers},0)"
    delta = f"INDEX({delta_values},1,{match})"
    comparator = f"INDEX({fcf_values},1,{match}-4)"
    return (
        f'=IFERROR(IF(OR($B$26="",COUNTIF({headers},$B$26)<>1,{match}<=4,'
        f'NOT(ISNUMBER({delta})),NOT(ISNUMBER({comparator})),{comparator}=0),"",'
        f'{delta}/ABS({comparator})),"")'
    )


def _apply_summary_formulas(ws: Any, enabled_formula_ids: set[str]) -> None:
    ws["A33"] = "GAAP EPS YoY % (latest vs LY quarter)"
    ws["C33"] = "%"
    formulas = {
        "summary_revenue_ttm": ("B27", _summary_valuation_value_formula(10), "#,##0.0;[Red]-#,##0.0"),
        "summary_revenue_yoy": ("B29", _summary_valuation_value_formula(11), "0.0%;[Red]-0.0%"),
        "summary_net_income_yoy": ("B31", _summary_valuation_value_formula(38), "0.0%;[Red]-0.0%"),
        "summary_gaap_eps": ("B32", _summary_valuation_value_formula(107), "$0.00;[Red]-$0.00"),
        "summary_gaap_eps_yoy": ("B33", _summary_valuation_value_formula(108), "0.0%;[Red]-0.0%"),
        "summary_free_cash_flow_ttm": ("B36", _summary_valuation_value_formula(49), "#,##0.0;[Red]-#,##0.0"),
        "summary_free_cash_flow_yoy": ("B37", _summary_free_cash_flow_yoy_formula(), "0.0%;[Red]-0.0%"),
        "summary_interest_coverage": ("B42", _summary_valuation_value_formula(88), "0.00x;[Red]-0.00x"),
    }
    for formula_id, (coordinate, formula, number_format) in formulas.items():
        if formula_id in enabled_formula_ids:
            _set_formula(ws[coordinate], formula, number_format)


def _apply_valuation_quarterly_formulas(ws: Any, enabled_formula_ids: set[str]) -> None:
    helper = VALUATION_HELPER_ROWS
    for column in range(FIRST_QUARTER_COLUMN, LAST_QUARTER_COLUMN + 1):
        col = get_column_letter(column)
        prior = get_column_letter(column - 4) if column - 4 >= FIRST_QUARTER_COLUMN else ""
        period_cell = f"{col}$6"

        formulas: dict[int, str] = {
            10: _history_ttm_sum(period_cell, "revenue", "$m"),
            11: _history_yoy_ratio(period_cell, "revenue", "$m"),
            13: _ratio(f"{col}{helper['gross_profit']}", f"{col}9"),
            14: _ratio(f"{col}32", f"{col}9"),
            15: _history_ttm_ratio(period_cell, "operating_income", "revenue", "$m"),
            19: _ratio(f"{col}18", f"{col}9"),
            20: _history_yoy_ratio(period_cell, "base_ebitda", "$m"),
            21: _history_ttm_sum(period_cell, "base_ebitda", "$m"),
            22: _ratio(f"{col}21", f"{col}10"),
            25: _history_ttm_sum(period_cell, "adjusted_ebitda", "$m"),
            26: _difference(f"{col}24", f"{col}18"),
            27: _ratio(f"{col}24", f"{col}9"),
            28: _history_yoy_ratio(period_cell, "adjusted_ebitda", "$m"),
            29: _ratio(f"{col}25", f"{col}10"),
            33: _ratio(f"{col}32", f"{col}9"),
            34: _history_ttm_sum(period_cell, "operating_income", "$m"),
            35: _ratio(f"{col}34", f"{col}10"),
            37: _ratio(f"{col}36", f"{col}9"),
            38: _history_yoy_ratio(period_cell, "net_income", "$m"),
            39: _history_ttm_sum(period_cell, "net_income", "$m"),
            40: _ratio(f"{col}39", f"{col}10"),
            45: _ratio(f"{col}44", f"{col}9"),
            46: _history_ttm_ratio(period_cell, "capital_expenditures", "revenue", "$m"),
            47: _difference(f"{col}43", f"{col}44"),
            48: _history_yoy_difference_of_differences(period_cell, "operating_cash_flow", "capital_expenditures", "$m"),
            49: _history_ttm_difference(period_cell, "operating_cash_flow", "capital_expenditures", "$m"),
            50: _history_ttm_yoy_difference_of_differences(period_cell, "operating_cash_flow", "capital_expenditures", "$m"),
            57: _ratio(f"{col}47", f"{col}9"),
            58: _ratio(f"{col}49", f"{col}10"),
            63: _history_ttm_sum(period_cell, "buybacks_cash", "$m"),
            64: _history_ttm_sum(period_cell, "dividends_cash", "$m"),
            65: _history_ttm_sum(period_cell, "acquisitions_cash", "$m"),
            66: _history_ttm_sum(period_cell, "debt_repayment", "$m"),
            67: _history_ttm_sum(period_cell, "debt_issuance", "$m"),
            73: _difference(f"{col}72", f"{col}70"),
            74: _qoq_difference(column, 73),
            75: _yoy_difference(prior, col, 73),
            77: f'=IF({col}73="","",-{col}73)',
            78: f'=IF(OR({col}70="",{col}71="",{col}72=""),"",{col}70+{col}71-{col}72)',
            80: f'=IF(OR({col}70="",{col}72="",{col}79=""),"",{col}72+{col}79-{col}70)',
            81: f'=IF(OR({col}70="",{col}71="",{col}72="",{col}79=""),"",{col}72+{col}79-{col}70-{col}71)',
            84: f'=IF({col}21="","",{col}21)',
            85: f'=IF({col}25="","",{col}25)',
            86: _ratio(f"{col}73", f"{col}84"),
            87: _ratio(f"{col}73", f"{col}85"),
            88: _history_ttm_ratio(period_cell, "operating_income", "interest_expense", "$m"),
            89: _history_ttm_ratio(period_cell, "base_ebitda", "interest_paid", "$m"),
            90: _ratio(f"{col}49", f"{col}21"),
            104: _history_point_difference(period_cell, "diluted_shares", "m shares", offset=-1),
            105: _history_point_difference(period_cell, "diluted_shares", "m shares", offset=-4),
            108: _history_yoy_ratio(period_cell, "eps", "$/share"),
            109: _history_ttm_sum(period_cell, "eps", "$/share"),
            111: _history_ttm_sum(period_cell, "adjusted_eps", "$/share"),
            113: _ratio(f"{col}{helper['total_equity']}", f"{col}103"),
            114: f'=IF(OR({col}{helper["total_equity"]}="",{col}{helper["goodwill"]}="",{col}{helper["intangibles"]}="",{col}103="",{col}103=0),"",({col}{helper["total_equity"]}-{col}{helper["goodwill"]}-{col}{helper["intangibles"]})/{col}103)',
            115: _ratio(f"{col}49", f"{col}102"),
            271: _history_ttm_sum(period_cell, "operating_cash_flow", "$m"),
        }
        formula_ids_by_row = {contract.row: contract.formula_id for contract in FORMULA_ROWS}
        for row, formula in formulas.items():
            if formula_ids_by_row[row] in enabled_formula_ids:
                _set_formula(ws.cell(row, column), formula, _number_format_for_row(row))


def _apply_balance_sheet_formulas(ws: Any, enabled_formula_ids: set[str]) -> None:
    for column in range(FIRST_QUARTER_COLUMN, LAST_QUARTER_COLUMN + 1):
        col = get_column_letter(column)
        prior = get_column_letter(column - 1) if column > 2 else ""
        yoy = get_column_letter(column - 4) if column > 5 else ""
        formulas = {
            11: f'=IF(OR({col}9="",{col}10=""),"",{col}9+{col}10)',
            12: _difference(f"{col}9", f"{prior}9") if prior else '=""',
            26: _ratio(f"{col}22", f"{col}25"),
            36: _difference(f"{col}18", f"{col}35"),
            37: _difference(f"{col}36", f"{prior}36") if prior else '=""',
            38: _ratio(f"{col}18", f"{col}35"),
            39: f'=IF(OR({col}18="",{col}15="",{col}35="",{col}35=0),"",({col}18-{col}15)/{col}35)',
            41: _difference(f"{col}40", f"{prior}40") if prior else '=""',
            51: _yoy_ratio(yoy, col, 15),
            52: _bs_sales_yoy_formula(col),
            53: f'=IF(OR({col}51="",{col}52=""),"",{col}51-{col}52)',
            54: f'=IF(OR({col}9="",{col}13="",{col}40=""),"",{col}9+{col}13-{col}40)',
            55: f'=IF(OR({col}34="",{col}42=""),"",{col}34+{col}42)',
            56: _yoy_ratio(yoy, col, 49),
        }
        formula_ids_by_row = {contract.row: contract.formula_id for contract in BS_QUARTERLY_FORMULA_ROWS}
        for row, formula in formulas.items():
            if formula_ids_by_row[row] not in enabled_formula_ids:
                continue
            number_format = "0.0%;[Red]-0.0%" if row in {26, 38, 39, 51, 52, 53, 56} else "#,##0.0;[Red]-#,##0.0"
            _set_formula(ws.cell(row, column), formula, number_format)


def _apply_valuation_input_outputs(ws: Any, enabled_formula_ids: set[str]) -> None:
    if not {"valuation_output_formulas", "valuation_scenario_formulas"} & enabled_formula_ids:
        return
    labels = {
        "B199": "EBITDA (base, TTM)",
        "B200": "Adjusted EBITDA (TTM)",
        "B202": "CFO TTM ($m)",
        "B204": "Diluted EPS (GAAP, TTM)",
        "B205": "Adjusted diluted EPS (TTM)",
        "B206": "Book value/share",
        "B207": "Tangible book value/share",
        "K206": "P/E (GAAP, TTM)",
        "K207": "P/E (adjusted, TTM)",
        "F210": "Enter a decimal percentage (for example, 10%); values above 100% are invalid.",
    }
    for coordinate, label in labels.items():
        ws[coordinate] = label

    denominator = 'IF(PerShareMode="Outstanding",Shares,IF(PerShareMode="Diluted",SharesDiluted,""))'
    output_formulas = {
        "N194": '=IF(OR(Price="",Shares=""),"",Price*Shares)',
        "N195": '=IF(OR(MarketCap="",NetDebt=""),"",MarketCap+NetDebt)',
        "N196": '=IF(OR(EV="",Adj_EBITDA="",Adj_EBITDA=0),"",EV/Adj_EBITDA)',
        "N197": '=IF(OR(EV="",Base_EBITDA="",Base_EBITDA=0),"",EV/Base_EBITDA)',
        "N198": '=IF(OR(FCF_TTM="",InterestPaid_TTM=""),"",FCF_TTM+InterestPaid_TTM)',
        "N199": '=IF(OR(FCFF_Proxy_TTM="",EV="",EV=0),"",FCFF_Proxy_TTM/EV)',
        "N200": '=IF(OR(FCF_TTM="",MarketCap="",MarketCap=0),"",FCF_TTM/MarketCap)',
        "N201": '=IF(OR(FCF_TTM="",Capex_TTM="",MaintCapexRatio="",RecurringCashCosts="",WCNormalization=""),"",FCF_TTM+(1-MaintCapexRatio)*Capex_TTM-RecurringCashCosts+WCNormalization)',
        "N202": '=IF(OR(OwnerEarnings_TTM="",EV="",EV=0),"",OwnerEarnings_TTM/EV)',
        "N203": f'=IF(OR(Target_EV_AdjEBITDA="",Adj_EBITDA="",NetDebt="",{denominator}="",{denominator}=0),"",(Target_EV_AdjEBITDA*Adj_EBITDA-NetDebt)/{denominator})',
        "N204": f'=IF(OR(Target_EV_EBITDA="",Base_EBITDA="",NetDebt="",{denominator}="",{denominator}=0),"",(Target_EV_EBITDA*Base_EBITDA-NetDebt)/{denominator})',
        "N205": f'=IF(OR(Target_EV_Yield="",Target_EV_Yield<=0,Target_EV_Yield>1,FCFF_Proxy_TTM="",NetDebt="",{denominator}="",{denominator}=0),"",(FCFF_Proxy_TTM/Target_EV_Yield-NetDebt)/{denominator})',
        "N206": '=IF(OR(Price="",EPS_TTM="",EPS_TTM=0),"",Price/EPS_TTM)',
        "N207": '=IF(OR(Price="",Adj_EPS_TTM="",Adj_EPS_TTM=0),"",Price/Adj_EPS_TTM)',
        "N208": '=IF(OR(EV="",Revenue_TTM="",Revenue_TTM=0),"",EV/Revenue_TTM)',
        "N209": '=IF(OR(Price="",BV_PerShare="",BV_PerShare=0),"",Price/BV_PerShare)',
        "N210": '=IF(OR(Price="",TBV_PerShare="",TBV_PerShare=0),"",Price/TBV_PerShare)',
    }
    if "valuation_output_formulas" in enabled_formula_ids:
        for coordinate, formula in output_formulas.items():
            number_format = "0.00x" if coordinate in {"N196", "N197", "N206", "N207", "N208", "N209", "N210"} else ("0.0%" if coordinate in {"N199", "N200", "N202"} else "#,##0.0")
            _set_formula(ws[coordinate], formula, number_format)

    _apply_valuation_scenario_inputs_and_outputs(ws, enabled_formula_ids)


def _apply_valuation_scenario_inputs_and_outputs(ws: Any, enabled_formula_ids: set[str]) -> None:
    _prepare_valuation_scenario_layout(ws)
    labels = {
        "B217": "Net income TTM ($m)",
        "B218": "Target EV/Revenue (x)",
        "B219": "Target P/E (x)",
        "B220": "DCF horizon (years)",
        "G218": "Starting FCFF ($m)",
        "G219": "Growth (projection period)",
        "G220": "Terminal growth",
        "G221": "WACC",
        "G225": "DCF sensitivity ($/share)",
        "G226": "WACC / terminal growth",
        "B235": "Scenario economics",
        "B236": "Scenario label",
        "B237": "Revenue growth",
        "B238": "Base EBITDA margin",
        "B239": "Adjusted EBITDA margin",
        "B240": "Tax rate",
        "H236": "Scenario horizon",
        "B241": "Scenario revenue ($m)",
        "B242": "Scenario EBITDA (base, $m)",
        "B243": "Scenario adjusted EBITDA ($m)",
        "B244": "Scenario FCF ($m)",
        "B246": "Explicit scenario bridge inputs",
        "B247": "Pre-tax earnings bridge ($m)",
        "B248": "Cash-interest change ($m)",
        "B249": "Capex change ($m, positive outflow)",
        "B250": "Working-capital adjustment ($m, positive source)",
        "B252": "Capital allocation inputs",
        "B253": "Buyback cash ($m)",
        "B254": "Buyback execution price ($/share)",
        "B255": "Share issuance (m shares)",
        "B256": "Debt paydown ($m)",
        "B258": "Scenario outputs",
        "B259": "Scenario shares (m)",
        "B260": "Scenario net debt ($m)",
        "B261": "Scenario EPS ($/share)",
        "H259": "Scenario enterprise value ($m)",
        "H260": "Scenario equity value ($m)",
        "H261": "Scenario implied price ($/share)",
        "L236": "Market cap ($m)",
        "L237": "Enterprise value ($m)",
        "L238": "Required revenue @ target EV/Revenue ($m)",
        "L239": "Required adjusted EBITDA @ target multiple ($m)",
        "L240": "Implied adjusted EBITDA margin",
        "L241": "Implied revenue growth",
        "L242": "Required FCFF @ target yield ($m)",
        "L243": "Required EPS @ target P/E ($/share)",
        "L244": "Implied terminal growth (perpetuity)",
        "M259": "Upside / downside",
        "M260": "Scenario P/E",
        "M261": "Scenario FCF yield",
        "Q259": "Scenario EV/Revenue",
        "Q260": "Scenario EV/EBITDA (base)",
        "Q261": "Scenario EV/Adjusted EBITDA",
    }
    for coordinate, label in labels.items():
        ws[coordinate] = label

    user_inputs = (
        "D218", "D219", "D220", "J218", "J219", "J220", "J221",
        "E236", "E237", "E238", "E239", "E240", "J236", "D247", "E247",
        "D248", "E248", "D249", "D250", "D253", "D254", "D255", "D256",
    )
    sensitivity_inputs = tuple(f"{column}226" for column in "HIJKL") + tuple(f"G{row}" for row in range(227, 235))
    for coordinate in ("D217",) + user_inputs + sensitivity_inputs:
        ws[coordinate] = None
        ws[coordinate].protection = Protection(locked=True)
    if "valuation_scenario_formulas" not in enabled_formula_ids:
        return

    denominator = _per_share_denominator_formula()
    dcf_ev = (
        "DCF_FCFF*(1+DCF_Growth)/(DCF_WACC-DCF_Growth)*"
        "(1-((1+DCF_Growth)/(1+DCF_WACC))^DCF_Horizon)+"
        "DCF_FCFF*(1+DCF_Growth)^DCF_Horizon*(1+DCF_TerminalGrowth)/"
        "(DCF_WACC-DCF_TerminalGrowth)/(1+DCF_WACC)^DCF_Horizon"
    )
    _set_formula(
        ws["J222"],
        '=IF(OR(DCF_FCFF="",DCF_FCFF<=0,DCF_Growth="",DCF_TerminalGrowth="",DCF_WACC="",DCF_Horizon="",DCF_Horizon<1,DCF_WACC<=DCF_Growth,DCF_WACC<=DCF_TerminalGrowth),"",'
        + dcf_ev
        + ")",
        "#,##0.0;[Red]-#,##0.0",
    )
    _set_formula(
        ws["J223"],
        f'=IF(OR(DCF_EV="",NetDebt="",{denominator}="",{denominator}<=0),"",(DCF_EV-NetDebt)/{denominator})',
        "$#,##0.00;[Red]-$#,##0.00",
    )
    _set_formula(ws["Q219"], '=IF(DCF_EV="","",DCF_EV)', "#,##0.0;[Red]-#,##0.0")
    _set_formula(ws["Q220"], '=IF(DCF_WACC="","",DCF_WACC)', "0.0%")
    _set_formula(ws["Q221"], '=IF(DCF_TerminalGrowth="","",DCF_TerminalGrowth)', "0.0%")

    for row in range(227, 235):
        for column in range(8, 13):
            coordinate = f"{get_column_letter(column)}{row}"
            wacc = f"$G${row}"
            terminal = f"${get_column_letter(column)}$226"
            grid_ev = (
                f"DCF_FCFF*(1+DCF_Growth)/({wacc}-DCF_Growth)*(1-((1+DCF_Growth)/(1+{wacc}))^DCF_Horizon)+"
                f"DCF_FCFF*(1+DCF_Growth)^DCF_Horizon*(1+{terminal})/({wacc}-{terminal})/(1+{wacc})^DCF_Horizon"
            )
            _set_formula(
                ws[coordinate],
                f'=IF(OR(DCF_FCFF="",DCF_FCFF<=0,DCF_Growth="",DCF_Horizon="",DCF_Horizon<1,{wacc}="",{terminal}="",{wacc}<=DCF_Growth,{wacc}<={terminal},NetDebt="",{denominator}="",{denominator}<=0),"",({grid_ev}-NetDebt)/{denominator})',
                "$#,##0.00;[Red]-$#,##0.00",
            )

    requirements = (
        'ScenarioHorizon="",ResolvedRevenueGrowth_Custom="",ScenarioBaseMargin="",ScenarioAdjustedMargin="",ScenarioTaxRate="",'
        'ScenarioPreTaxBridge="",ScenarioTaxTreatment="",ScenarioCashInterestChange="",'
        'ScenarioInterestTaxTreatment="",ScenarioCapexChange="",ScenarioWCAdjustment="",'
        'ScenarioBuybackCash="",ScenarioShareIssuance="",ScenarioDebtPaydown="",Revenue_TTM="",'
        f'FCF_TTM="",NetIncome_TTM="",NetDebt="",{denominator}="",'
        'AND(ScenarioBuybackCash<>0,OR(ScenarioBuybackPrice="",ScenarioBuybackPrice<=0))'
    )
    _set_formula(ws["F236"], f'=IF(ScenarioProfile="","",IF(OR({requirements}),"Incomplete","Ready"))', "General")
    _set_formula(ws["E241"], '=IF(OR(Revenue_TTM="",ResolvedRevenueGrowth_Custom=""),"",Revenue_TTM*(1+ResolvedRevenueGrowth_Custom))', "#,##0.0;[Red]-#,##0.0")
    _set_formula(ws["E242"], '=IF(OR(ScenarioRevenue="",ScenarioBaseMargin=""),"",ScenarioRevenue*ScenarioBaseMargin)', "#,##0.0;[Red]-#,##0.0")
    _set_formula(ws["E243"], '=IF(OR(ScenarioRevenue="",ScenarioAdjustedMargin=""),"",ScenarioRevenue*ScenarioAdjustedMargin)', "#,##0.0;[Red]-#,##0.0")

    fcf_bridge = _scenario_after_tax_bridge_formula(cash_view=True)
    eps_bridge = _scenario_after_tax_bridge_formula(cash_view=False)
    interest_bridge = _scenario_interest_after_tax_formula()
    _set_formula(
        ws["E244"],
        f'=IF(OR(FCF_TTM="",ScenarioPreTaxBridge="",ScenarioTaxTreatment="",ScenarioTaxRate="",ScenarioCashInterestChange="",ScenarioInterestTaxTreatment="",ScenarioCapexChange="",ScenarioWCAdjustment=""),"",FCF_TTM+{fcf_bridge}-{interest_bridge}-ScenarioCapexChange+ScenarioWCAdjustment)',
        "#,##0.0;[Red]-#,##0.0",
    )
    _set_formula(
        ws["E259"],
        f'=IF(OR({denominator}="",ScenarioBuybackCash="",ScenarioShareIssuance="",AND(ScenarioBuybackCash<>0,OR(ScenarioBuybackPrice="",ScenarioBuybackPrice<=0))),"",IF({denominator}-IF(ScenarioBuybackCash=0,0,ScenarioBuybackCash/ScenarioBuybackPrice)+ScenarioShareIssuance<=0,"",{denominator}-IF(ScenarioBuybackCash=0,0,ScenarioBuybackCash/ScenarioBuybackPrice)+ScenarioShareIssuance))',
        "#,##0.000;[Red]-#,##0.000",
    )
    _set_formula(ws["E260"], '=IF(OR(NetDebt="",ScenarioBuybackCash="",ScenarioDebtPaydown=""),"",NetDebt+ScenarioBuybackCash-ScenarioDebtPaydown)', "#,##0.0;[Red]-#,##0.0")
    _set_formula(
        ws["E261"],
        f'=IF(OR(NetIncome_TTM="",ScenarioShares="",ScenarioShares<=0,ScenarioPreTaxBridge="",ScenarioTaxTreatment="",ScenarioTaxRate="",ScenarioCashInterestChange="",ScenarioInterestTaxTreatment=""),"",(NetIncome_TTM+{eps_bridge}-{interest_bridge})/ScenarioShares)',
        "$0.00;[Red]-$0.00",
    )
    _set_formula(ws["J259"], '=IF(OR(Target_EV_AdjEBITDA="",ScenarioAdjustedEBITDA=""),"",Target_EV_AdjEBITDA*ScenarioAdjustedEBITDA)', "#,##0.0;[Red]-#,##0.0")
    _set_formula(ws["J260"], '=IF(OR(ScenarioEV="",ScenarioNetDebt=""),"",ScenarioEV-ScenarioNetDebt)', "#,##0.0;[Red]-#,##0.0")
    _set_formula(ws["J261"], '=IF(OR(ScenarioEquityValue="",ScenarioShares="",ScenarioShares<=0),"",ScenarioEquityValue/ScenarioShares)', "$#,##0.00;[Red]-$#,##0.00")
    _set_formula(ws["N259"], '=IF(OR(ScenarioImpliedPrice="",Price="",Price<=0),"",ScenarioImpliedPrice/Price-1)', "0.0%;[Red]-0.0%")
    _set_formula(ws["N260"], '=IF(OR(ScenarioImpliedPrice="",ScenarioEPS="",ScenarioEPS=0),"",ScenarioImpliedPrice/ScenarioEPS)', "0.00x")
    _set_formula(
        ws["N261"],
        '=IF(OR(NOT(ISNUMBER(ScenarioFCF)),NOT(ISNUMBER(ScenarioImpliedPrice)),NOT(ISNUMBER(ScenarioShares)),ScenarioImpliedPrice<=0,ScenarioShares<=0),"",ScenarioFCF/(ScenarioImpliedPrice*ScenarioShares))',
        "0.0%",
    )
    _set_formula(ws["R259"], '=IF(OR(ScenarioEV="",ScenarioRevenue="",ScenarioRevenue=0),"",ScenarioEV/ScenarioRevenue)', "0.00x")
    _set_formula(ws["R260"], '=IF(OR(ScenarioEV="",ScenarioBaseEBITDA="",ScenarioBaseEBITDA=0),"",ScenarioEV/ScenarioBaseEBITDA)', "0.00x")
    _set_formula(ws["R261"], '=IF(OR(ScenarioEV="",ScenarioAdjustedEBITDA="",ScenarioAdjustedEBITDA=0),"",ScenarioEV/ScenarioAdjustedEBITDA)', "0.00x")

    market_formulas = {
        "N236": f'=IF(OR(Price="",{denominator}=""),"",Price*{denominator})',
        "N237": '=IF(OR(N236="",NetDebt=""),"",N236+NetDebt)',
        "N238": '=IF(OR(N237="",Target_EV_Revenue="",Target_EV_Revenue<=0),"",N237/Target_EV_Revenue)',
        "N239": '=IF(OR(N237="",Target_EV_AdjEBITDA="",Target_EV_AdjEBITDA<=0),"",N237/Target_EV_AdjEBITDA)',
        "N240": '=IF(OR(N238="",N239="",N238=0),"",N239/N238)',
        "N241": '=IF(OR(N238="",Revenue_TTM="",Revenue_TTM=0),"",N238/Revenue_TTM-1)',
        "N242": '=IF(OR(N237="",Target_EV_Yield="",Target_EV_Yield<=0),"",N237*Target_EV_Yield)',
        "N243": '=IF(OR(Price="",Target_PE="",Target_PE<=0),"",Price/Target_PE)',
        "N244": '=IF(OR(NOT(ISNUMBER(N237)),NOT(ISNUMBER(DCF_WACC)),NOT(ISNUMBER(DCF_FCFF)),N237<=0,DCF_WACC<=0),"",IF(N237+DCF_FCFF=0,"",(N237*DCF_WACC-DCF_FCFF)/(N237+DCF_FCFF)))',
    }
    for coordinate, formula in market_formulas.items():
        number_format = "0.0%" if coordinate in {"N240", "N241", "N244"} else ("$0.00" if coordinate == "N243" else "#,##0.0")
        _set_formula(ws[coordinate], formula, number_format)


def _scenario_after_tax_bridge_formula(*, cash_view: bool) -> str:
    if cash_view:
        full_value_treatments = 'OR(ScenarioTaxTreatment="non_taxable",ScenarioTaxTreatment="non_taxable_credit",ScenarioTaxTreatment="cash_only",ScenarioTaxTreatment="no_eps_impact")'
        zero_value = "ScenarioPreTaxBridge"
    else:
        full_value_treatments = 'OR(ScenarioTaxTreatment="non_taxable",ScenarioTaxTreatment="non_taxable_credit")'
        zero_value = "0"
    return (
        'IF(ScenarioTaxTreatment="taxable",ScenarioPreTaxBridge*(1-ScenarioTaxRate),'
        f'IF({full_value_treatments},ScenarioPreTaxBridge,'
        f'IF(OR(ScenarioTaxTreatment="cash_only",ScenarioTaxTreatment="no_eps_impact"),{zero_value},"")))'
    )


def _scenario_interest_after_tax_formula() -> str:
    return 'IF(ScenarioInterestTaxTreatment="taxable",ScenarioCashInterestChange*(1-ScenarioTaxRate),IF(ScenarioInterestTaxTreatment="non_taxable",ScenarioCashInterestChange,""))'


def _per_share_denominator_formula() -> str:
    return 'IF(PerShareMode="Outstanding",Shares,IF(PerShareMode="Diluted",SharesDiluted,""))'


def _apply_valuation_sidecar_outputs(ws: Any, enabled_formula_ids: set[str]) -> None:
    if "valuation_sidecar_formulas" not in enabled_formula_ids:
        return
    labels = {
        64: "Adjusted EBITDA TTM",
        65: "FCF TTM",
        66: "Diluted EPS (GAAP, TTM)",
        67: "EV @ target EV/Adjusted EBITDA",
        68: "Equity value @ target EV/Adjusted EBITDA",
        69: "EV @ target EV/EBITDA",
        70: "Equity value @ target EV/EBITDA",
        72: "Per-share valuation outputs",
        73: "Value/share @ target EV/Adjusted EBITDA",
        74: "Value/share @ target EV/EBITDA",
        75: "Value/share @ target FCFF yield",
    }
    formulas = {
        64: '=IF(Adj_EBITDA="","",Adj_EBITDA)',
        65: '=IF(FCF_TTM="","",FCF_TTM)',
        66: '=IF(EPS_TTM="","",EPS_TTM)',
        67: '=IF(OR(Adj_EBITDA="",Target_EV_AdjEBITDA=""),"",Adj_EBITDA*Target_EV_AdjEBITDA)',
        68: '=IF(OR(U67="",NetDebt=""),"",U67-NetDebt)',
        69: '=IF(OR(Base_EBITDA="",Target_EV_EBITDA=""),"",Base_EBITDA*Target_EV_EBITDA)',
        70: '=IF(OR(U69="",NetDebt=""),"",U69-NetDebt)',
        72: '=TEXT(MIN(U73,U74,U75),"$0")&"-"&TEXT(MAX(U73,U74,U75),"$0")',
        73: '=IF(EqShare_Target_Adj="","",EqShare_Target_Adj)',
        74: '=IF(EqShare_Target_EV="","",EqShare_Target_EV)',
        75: '=IF(EqShare_Target_Yield="","",EqShare_Target_Yield)',
    }
    for row, label in labels.items():
        ws.cell(row, 15).value = label
        if row in formulas:
            _set_formula(ws.cell(row, 21), formulas[row], "$#,##0.00;[Red]-$#,##0.00" if row >= 73 else "#,##0.0;[Red]-#,##0.0")


def _prepare_valuation_scenario_layout(ws: Any) -> None:
    from openpyxl.utils.cell import range_boundaries

    exact_surfaces = (
        "B217:E220",
        "G218:J223",
        "L236:O244",
        "H236:J236",
        "B247:K250",
        "B253:D256",
        "F259:K261",
        "N259:N261",
        "R259:R261",
    )
    owned = [range_boundaries(target) for target in exact_surfaces]
    for merged in tuple(ws.merged_cells.ranges):
        bounds = (merged.min_col, merged.min_row, merged.max_col, merged.max_row)
        if any(_bounds_overlap(bounds, target) for target in owned):
            ws.unmerge_cells(str(merged))


def _apply_investment_case_scenario_formulas(ws: Any, enabled_formula_ids: set[str]) -> None:
    _prepare_investment_case_scenario_layout(ws)
    labels = {
        13: "Typed Scenario Inputs",
        14: "Input / assumption",
        15: "Current share price",
        16: "Base shares (selected denominator)",
        17: "Net debt",
        18: "Revenue TTM",
        19: "Net income TTM",
        20: "EBITDA (base, TTM)",
        21: "Adjusted EBITDA (TTM)",
        22: "FCF TTM",
        23: "Fiscal horizon",
        24: "Revenue growth",
        25: "Base EBITDA margin",
        26: "Adjusted EBITDA margin",
        27: "Pre-tax earnings bridge ($m)",
        28: "Tax rate",
        29: "Earnings-bridge tax treatment",
        30: "Cash-interest change ($m)",
        31: "Cash-interest tax treatment",
        32: "Capex change ($m, positive outflow)",
        33: "Working-capital adjustment ($m, positive source)",
        34: "Buyback cash ($m)",
        35: "Buyback execution price ($/share)",
        36: "Share issuance (m shares)",
        37: "Debt paydown ($m)",
        38: "Target EV/Adjusted EBITDA (x)",
        39: "Target EV/EBITDA (base, x)",
        40: "Target EV/Revenue (x)",
        41: "Target P/E (x)",
        42: "Target FCF yield",
        47: "Typed Scenario Driver Bridge",
        48: "Only explicit typed assumptions affect scenario economics; source guidance remains separate until selected.",
        49: "Bridge output",
        50: "After-tax earnings bridge ($m)",
        51: "After-tax cash-interest change ($m)",
        52: "Net FCF bridge ($m)",
        53: "Net share-count change (m)",
        55: "Scenario output",
        56: "Scenario EPS ($/share)",
        57: "Scenario adjusted EBITDA ($m)",
        58: "Scenario FCF ($m)",
        60: "What Market Is Pricing",
        61: "Metric",
        62: "Market price",
        63: "Market capitalization ($m)",
        64: "Enterprise value ($m)",
        65: "Required revenue @ target EV/Revenue ($m)",
        66: "Required adjusted EBITDA @ target multiple ($m)",
        67: "Required FCFF @ target yield ($m)",
        68: "Implied terminal growth (perpetuity)",
        83: "Bear / Base / Bull Scenario",
        84: "Scenario",
        85: "Bear",
        86: "Base",
        87: "Bull",
        159: "P/E Sensitivity ($/share)",
        160: "EPS / P-E multiple",
        170: "EV/Adjusted EBITDA Sensitivity ($/share)",
        171: "Adjusted EBITDA / multiple",
        176: "FCF Yield Sensitivity ($/share)",
        177: "FCF / yield",
    }
    for row, label in labels.items():
        ws.cell(row, 1).value = label
    for column, scenario in zip(range(2, 5), ("Bear", "Base", "Bull"), strict=True):
        ws.cell(14, column).value = scenario
        ws.cell(49, column).value = scenario
        ws.cell(55, column).value = scenario
    for column, label in enumerate(
        ("Revenue", "Base EBITDA", "Adjusted EBITDA", "FCF", "EPS", "Shares", "Net debt", "Implied price", "Upside / downside"),
        start=2,
    ):
        ws.cell(84, column).value = label

    for range_ref in user_input_surface_targets("{ticker}_Investment_Case"):
        _clear_range(ws, range_ref)

    if "investment_case_scenario_formulas" in enabled_formula_ids:
        actuals = {
            15: ("Price", "$#,##0.00;[Red]-$#,##0.00"),
            16: (_per_share_denominator_formula(), "#,##0.000;[Red]-#,##0.000"),
            17: ("NetDebt", "#,##0.0;[Red]-#,##0.0"),
            18: ("Revenue_TTM", "#,##0.0;[Red]-#,##0.0"),
            19: ("NetIncome_TTM", "#,##0.0;[Red]-#,##0.0"),
            20: ("Base_EBITDA", "#,##0.0;[Red]-#,##0.0"),
            21: ("Adj_EBITDA", "#,##0.0;[Red]-#,##0.0"),
            22: ("FCF_TTM", "#,##0.0;[Red]-#,##0.0"),
        }
        for row, (name_or_formula, number_format) in actuals.items():
            for column in range(2, 5):
                expression = name_or_formula
                _set_formula(ws.cell(row, column), f'=IF({expression}="","",{expression})', number_format)

        for input_column, output_row, resolved_growth_name in zip(
            "BCD",
            range(85, 88),
            ("ResolvedRevenueGrowth_Bear", "ResolvedRevenueGrowth_Base", "ResolvedRevenueGrowth_Bull"),
            strict=True,
        ):
            cash_bridge = _cell_scenario_bridge_formula(input_column, cash_view=True)
            eps_bridge = _cell_scenario_bridge_formula(input_column, cash_view=False)
            interest_bridge = _cell_interest_bridge_formula(input_column)
            _set_formula(ws[f"{input_column}50"], f'=IF(OR({input_column}27="",{input_column}28="",{input_column}29=""),"",{eps_bridge})', "#,##0.0;[Red]-#,##0.0")
            _set_formula(ws[f"{input_column}51"], f'=IF(OR({input_column}28="",{input_column}30="",{input_column}31=""),"",{interest_bridge})', "#,##0.0;[Red]-#,##0.0")
            _set_formula(ws[f"{input_column}52"], f'=IF(OR({input_column}27="",{input_column}28="",{input_column}29="",{input_column}30="",{input_column}31="",{input_column}32="",{input_column}33=""),"",{cash_bridge}-{interest_bridge}-{input_column}32+{input_column}33)', "#,##0.0;[Red]-#,##0.0")
            _set_formula(ws[f"{input_column}53"], f'=IF(OR({input_column}34="",{input_column}36="",AND({input_column}34<>0,OR({input_column}35="",{input_column}35<=0))),"",-IF({input_column}34=0,0,{input_column}34/{input_column}35)+{input_column}36)', "#,##0.000;[Red]-#,##0.000")

            _set_formula(ws[f"B{output_row}"], f'=IF(OR({input_column}18="",{resolved_growth_name}=""),"",{input_column}18*(1+{resolved_growth_name}))', "#,##0.0;[Red]-#,##0.0")
            _set_formula(ws[f"C{output_row}"], f'=IF(OR(B{output_row}="",{input_column}25=""),"",B{output_row}*{input_column}25)', "#,##0.0;[Red]-#,##0.0")
            _set_formula(ws[f"D{output_row}"], f'=IF(OR(B{output_row}="",{input_column}26=""),"",B{output_row}*{input_column}26)', "#,##0.0;[Red]-#,##0.0")
            _set_formula(ws[f"E{output_row}"], f'=IF(OR({input_column}22="",{input_column}52=""),"",{input_column}22+{input_column}52)', "#,##0.0;[Red]-#,##0.0")
            _set_formula(ws[f"G{output_row}"], f'=IF(OR({input_column}16="",{input_column}53=""),"",IF({input_column}16+{input_column}53<=0,"",{input_column}16+{input_column}53))', "#,##0.000;[Red]-#,##0.000")
            _set_formula(ws[f"H{output_row}"], f'=IF(OR({input_column}17="",{input_column}34="",{input_column}37=""),"",{input_column}17+{input_column}34-{input_column}37)', "#,##0.0;[Red]-#,##0.0")
            _set_formula(ws[f"F{output_row}"], f'=IF(OR({input_column}19="",{input_column}50="",{input_column}51="",G{output_row}="",G{output_row}<=0),"",({input_column}19+{input_column}50-{input_column}51)/G{output_row})', "$0.00;[Red]-$0.00")
            _set_formula(ws[f"I{output_row}"], f'=IF(OR(D{output_row}="",{input_column}38="",H{output_row}="",G{output_row}="",G{output_row}<=0),"",(D{output_row}*{input_column}38-H{output_row})/G{output_row})', "$#,##0.00;[Red]-$#,##0.00")
            _set_formula(ws[f"J{output_row}"], f'=IF(OR(I{output_row}="",{input_column}15="",{input_column}15<=0),"",I{output_row}/{input_column}15-1)', "0.0%;[Red]-0.0%")

        for column, output_row in zip("BCD", range(85, 88), strict=True):
            _set_formula(ws[f"{column}56"], f'=IF(F{output_row}="","",F{output_row})', "$0.00;[Red]-$0.00")
            _set_formula(ws[f"{column}57"], f'=IF(D{output_row}="","",D{output_row})', "#,##0.0;[Red]-#,##0.0")
            _set_formula(ws[f"{column}58"], f'=IF(E{output_row}="","",E{output_row})', "#,##0.0;[Red]-#,##0.0")

        market_links = {
            62: ("Price", "$#,##0.00;[Red]-$#,##0.00"),
            63: ("N236", "#,##0.0;[Red]-#,##0.0"),
            64: ("N237", "#,##0.0;[Red]-#,##0.0"),
            65: ("N238", "#,##0.0;[Red]-#,##0.0"),
            66: ("N239", "#,##0.0;[Red]-#,##0.0"),
            67: ("N242", "#,##0.0;[Red]-#,##0.0"),
            68: ("N244", "0.0%;[Red]-0.0%"),
        }
        for row, (source, number_format) in market_links.items():
            if row == 68:
                formula = f'=IF(ISNUMBER(\'Valuation\'!{source}),\'Valuation\'!{source},"")'
            else:
                formula = f"=IF('Valuation'!{source}=\"\",\"\",'Valuation'!{source})" if source.startswith("N") else f'=IF({source}="","",{source})'
            _set_formula(ws[f"B{row}"], formula, number_format)

    if "investment_case_sensitivity_formulas" in enabled_formula_ids:
        denominator = _per_share_denominator_formula()
        for row in range(161, 164):
            for column in range(2, 5):
                col = get_column_letter(column)
                _set_formula(ws.cell(row, column), f'=IF(OR($A{row}="",{col}$160=""),"",$A{row}*{col}$160)', "$#,##0.00;[Red]-$#,##0.00")
        for row in range(172, 175):
            for column in range(2, 5):
                col = get_column_letter(column)
                _set_formula(ws.cell(row, column), f'=IF(OR($A{row}="",{col}$171="",NetDebt="",{denominator}="",{denominator}<=0),"",($A{row}*{col}$171-NetDebt)/{denominator})', "$#,##0.00;[Red]-$#,##0.00")
        for row in range(178, 181):
            for column in range(2, 5):
                col = get_column_letter(column)
                _set_formula(ws.cell(row, column), f'=IF(OR($A{row}="",{col}$177="",{col}$177<=0,{denominator}="",{denominator}<=0),"",$A{row}/{col}$177/{denominator})', "$#,##0.00;[Red]-$#,##0.00")


def _apply_valuation_summary_formulas(ws: Any, enabled_formula_ids: set[str]) -> None:
    if "valuation_summary_formulas" not in enabled_formula_ids:
        return

    projections = (
        ("price", "Price", "$/share", "$#,##0.00;[Red]-$#,##0.00", "Observed or explicit user price."),
        ("market_cap", "MarketCap", "$m", "#,##0.0;[Red]-#,##0.0", "Price multiplied by the selected share denominator."),
        ("enterprise_value", "EV", "$m", "#,##0.0;[Red]-#,##0.0", "Market capitalization plus available net debt."),
        ("implied_ev_adjusted_ebitda", "Implied_EV_AdjEBITDA", "x", "0.00x", "Enterprise value divided by adjusted EBITDA TTM."),
        ("implied_ev_base_ebitda", "Implied_EV_EBITDA", "x", "0.00x", "Enterprise value divided by base EBITDA TTM."),
        ("implied_ev_revenue", "'Valuation'!N208", "x", "0.00x", "Enterprise value divided by revenue TTM."),
        ("equity_fcf_yield", "Equity_FCF_Yield", "%", "0.0%", "FCF TTM divided by market capitalization."),
        ("dcf_enterprise_value", "DCF_EV", "$m", "#,##0.0;[Red]-#,##0.0", "DCF enterprise value from explicit user assumptions."),
        ("dcf_implied_price", "DCF_ImpliedPrice", "$/share", "$#,##0.00;[Red]-$#,##0.00", "DCF equity value divided by the selected share denominator."),
        ("scenario_revenue", "ScenarioRevenue", "$m", "#,##0.0;[Red]-#,##0.0", "Revenue under the selected typed scenario."),
        ("scenario_base_ebitda", "ScenarioBaseEBITDA", "$m", "#,##0.0;[Red]-#,##0.0", "Base EBITDA under the selected typed scenario."),
        ("scenario_adjusted_ebitda", "ScenarioAdjustedEBITDA", "$m", "#,##0.0;[Red]-#,##0.0", "Adjusted EBITDA under the selected typed scenario."),
        ("scenario_fcf", "ScenarioFCF", "$m", "#,##0.0;[Red]-#,##0.0", "FCF under explicit bridge and cash classifications."),
        ("scenario_shares", "ScenarioShares", "m shares", "#,##0.000;[Red]-#,##0.000", "Selected base shares less repurchased shares plus issuance."),
        ("scenario_net_debt", "ScenarioNetDebt", "$m", "#,##0.0;[Red]-#,##0.0", "Available net debt plus buyback cash less debt paydown."),
        ("scenario_eps", "ScenarioEPS", "$/share", "$0.00;[Red]-$0.00", "Scenario net income divided by scenario shares."),
        ("scenario_enterprise_value", "ScenarioEV", "$m", "#,##0.0;[Red]-#,##0.0", "Scenario adjusted EBITDA multiplied by the explicit target multiple."),
        ("scenario_implied_price", "ScenarioImpliedPrice", "$/share", "$#,##0.00;[Red]-$#,##0.00", "Scenario equity value divided by scenario shares."),
        ("scenario_upside", "ScenarioUpside", "%", "0.0%;[Red]-0.0%", "Scenario implied price relative to current price."),
    )
    for row, (metric, expression, unit, number_format, definition) in enumerate(projections, start=2):
        ws.cell(row, 1).value = metric
        _set_formula(ws.cell(row, 2), f'=IF({expression}="","",{expression})', number_format)
        ws.cell(row, 3).value = unit
        _set_formula(ws.cell(row, 4), '=IF(AsOfQuarter="","",AsOfQuarter)', "General")
        ws.cell(row, 5).value = f"formula:{expression}"
        _set_formula(ws.cell(row, 6), f'=IF(B{row}="","unavailable","calculated")', "General")
        ws.cell(row, 7).value = definition


def _apply_valuation_grid_formulas(ws: Any, enabled_formula_ids: set[str]) -> None:
    if "valuation_grid_formulas" not in enabled_formula_ids:
        return

    row = 2
    for valuation_row in range(227, 235):
        for valuation_column in range(8, 13):
            column = get_column_letter(valuation_column)
            _set_formula(ws.cell(row, 1), '="dcf"', "General")
            _set_formula(ws.cell(row, 2), f'=IF(\'Valuation\'!$G${valuation_row}="","",\'Valuation\'!$G${valuation_row})', "0.0%")
            _set_formula(ws.cell(row, 3), f'=IF(\'Valuation\'!${column}$226="","",\'Valuation\'!${column}$226)', "0.0%")
            ws.cell(row, 4).value = "implied_price"
            _set_formula(ws.cell(row, 5), f'=IF(\'Valuation\'!${column}${valuation_row}="","",\'Valuation\'!${column}${valuation_row})', "$#,##0.00;[Red]-$#,##0.00")
            _set_formula(ws.cell(row, 6), f'=IF(E{row}="","unavailable","calculated")', "General")
            row += 1

    _set_formula(ws.cell(row, 1), '=IF(ScenarioProfile="","",ScenarioProfile)', "General")
    _set_formula(ws.cell(row, 2), '=IF(ScenarioProfile="","","selected_scenario")', "General")
    _set_formula(ws.cell(row, 3), '=IF(AsOfQuarter="","",AsOfQuarter)', "General")
    ws.cell(row, 4).value = "scenario_implied_price"
    _set_formula(ws.cell(row, 5), '=IF(ScenarioImpliedPrice="","",ScenarioImpliedPrice)', "$#,##0.00;[Red]-$#,##0.00")
    _set_formula(ws.cell(row, 6), f'=IF(E{row}="","unavailable","calculated")', "General")


def _cell_scenario_bridge_formula(column: str, *, cash_view: bool) -> str:
    if cash_view:
        full_treatments = f'OR({column}29="non_taxable",{column}29="non_taxable_credit",{column}29="cash_only",{column}29="no_eps_impact")'
        zero_value = f"{column}27"
    else:
        full_treatments = f'OR({column}29="non_taxable",{column}29="non_taxable_credit")'
        zero_value = "0"
    return (
        f'IF({column}29="taxable",{column}27*(1-{column}28),'
        f'IF({full_treatments},{column}27,'
        f'IF(OR({column}29="cash_only",{column}29="no_eps_impact"),{zero_value},"")))'
    )


def _cell_interest_bridge_formula(column: str) -> str:
    return f'IF({column}31="taxable",{column}30*(1-{column}28),IF({column}31="non_taxable",{column}30,""))'


def _apply_scenario_defined_names(workbook: Any, enabled_formula_ids: set[str]) -> None:
    obsolete = {
        "ScenarioMargin", "ScenarioBuyback", "ScenarioRefiNorm", "ScenarioOwnerEarnings",
        "ScenarioAdjEBITDA", "ScenarioEqShare_EVAdj", "ScenarioEqShare_Yield",
    }
    for name in obsolete:
        if name in workbook.defined_names:
            del workbook.defined_names[name]
    route_names = {
        "ResolvedRevenueGrowth_Bear": "H2",
        "ResolvedRevenueGrowth_Base": "I2",
        "ResolvedRevenueGrowth_Bull": "J2",
        "ResolvedRevenueGrowth_Custom": "K2",
    }
    for name in route_names:
        if name in workbook.defined_names:
            del workbook.defined_names[name]
    if "scenario_revenue_route_formulas" in enabled_formula_ids and "Valuation_Summary" in workbook.sheetnames:
        route_sheet = quote_sheetname("Valuation_Summary")
        for name, coordinate in route_names.items():
            workbook.defined_names.add(DefinedName(name, attr_text=f"{route_sheet}!${coordinate[0]}${coordinate[1:]}"))
    if "valuation_scenario_formulas" not in enabled_formula_ids:
        return
    cells = {
        "NetIncome_TTM": "D217",
        "Target_EV_Revenue": "D218",
        "Target_PE": "D219",
        "DCF_Horizon": "D220",
        "DCF_FCFF": "J218",
        "DCF_Growth": "J219",
        "DCF_TerminalGrowth": "J220",
        "DCF_WACC": "J221",
        "DCF_EV": "J222",
        "DCF_ImpliedPrice": "J223",
        "ScenarioProfile": "E236",
        "ScenarioHorizon": "J236",
        "ScenarioGrowth": "E237",
        "ScenarioBaseMargin": "E238",
        "ScenarioAdjustedMargin": "E239",
        "ScenarioTaxRate": "E240",
        "ScenarioPreTaxBridge": "D247",
        "ScenarioTaxTreatment": "E247",
        "ScenarioCashInterestChange": "D248",
        "ScenarioInterestTaxTreatment": "E248",
        "ScenarioCapexChange": "D249",
        "ScenarioWCAdjustment": "D250",
        "ScenarioBuybackCash": "D253",
        "ScenarioBuybackPrice": "D254",
        "ScenarioShareIssuance": "D255",
        "ScenarioDebtPaydown": "D256",
        "ScenarioRevenue": "E241",
        "ScenarioBaseEBITDA": "E242",
        "ScenarioAdjustedEBITDA": "E243",
        "ScenarioFCF": "E244",
        "ScenarioShares": "E259",
        "ScenarioNetDebt": "E260",
        "ScenarioEPS": "E261",
        "ScenarioEV": "J259",
        "ScenarioEquityValue": "J260",
        "ScenarioImpliedPrice": "J261",
        "ScenarioUpside": "N259",
        "ScenarioPE": "N260",
        "ScenarioFCFYield": "N261",
        "ScenarioEVRevenue": "R259",
        "ScenarioEVBaseEBITDA": "R260",
        "ScenarioEVAdjustedEBITDA": "R261",
    }
    sheet = quote_sheetname("Valuation")
    for name, coordinate in cells.items():
        if name in workbook.defined_names:
            del workbook.defined_names[name]
        workbook.defined_names.add(DefinedName(name, attr_text=f"{sheet}!${coordinate[0]}${coordinate[1:]}"))


def _add_validation(
    ws: Any,
    validation_type: str,
    target: str,
    *,
    operator: str | None = None,
    formula1: str | None = None,
    formula2: str | None = None,
) -> None:
    validation = DataValidation(
        type=validation_type,
        operator=operator,
        formula1=formula1,
        formula2=formula2,
        allow_blank=True,
    )
    validation.errorTitle = "Invalid scenario input"
    validation.error = "Enter a value compatible with the declared scenario contract."
    validation.errorStyle = "stop"
    validation.showErrorMessage = True
    ws.add_data_validation(validation)
    validation.add(target)


def _remove_data_validations_overlapping(ws: Any, targets: Collection[str]) -> None:
    from openpyxl.utils.cell import range_boundaries

    owned = [range_boundaries(target) for target in targets]
    retained = []
    for validation in ws.data_validations.dataValidation:
        ranges = [range_boundaries(str(cell_range)) for cell_range in validation.ranges.ranges]
        if any(_bounds_overlap(candidate, target) for candidate in ranges for target in owned):
            continue
        retained.append(validation)
    ws.data_validations.dataValidation = retained


def _bounds_overlap(left: tuple[int, int, int, int], right: tuple[int, int, int, int]) -> bool:
    return not (left[2] < right[0] or right[2] < left[0] or left[3] < right[1] or right[3] < left[1])


def _clear_range(ws: Any, range_ref: str) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.protection = Protection(locked=True)


def _prepare_investment_case_scenario_layout(ws: Any) -> None:
    from openpyxl.utils.cell import range_boundaries

    exact_surfaces = (
        "B15:D42",
        "B50:D53",
        "B55:D58",
        "B62:B68",
        "B84:J87",
        "B160:D160",
        "A161:D163",
        "B171:D171",
        "A172:D174",
        "B177:D177",
        "A178:D180",
    )
    owned = [range_boundaries(target) for target in exact_surfaces]
    for merged in tuple(ws.merged_cells.ranges):
        bounds = (merged.min_col, merged.min_row, merged.max_col, merged.max_row)
        if any(_bounds_overlap(bounds, target) for target in owned):
            ws.unmerge_cells(str(merged))


def _apply_hidden_value_support_formulas(workbook: Any, enabled_formula_ids: set[str]) -> None:
    formula_ids = {
        "hidden_value_recompute_detail_formulas",
        "hidden_value_recompute_candidate_formulas",
        "hidden_value_audit_parity_formulas",
        "hidden_value_valuation_state_count_formulas",
    }
    if "Valuation" in workbook.sheetnames:
        workbook["Valuation"]["AI139"].value = None
    _remove_hidden_value_defined_names(workbook)
    active = formula_ids & enabled_formula_ids
    if active and active != formula_ids:
        raise ValueError("Hidden Value workbook recomputation formula families must activate together.")
    if not active:
        return
    required_sheets = {"Hidden_Value_Base", "Hidden_Value_Audit", "Hidden_Value_Recompute", "Hidden_Value_Flags"}
    missing = sorted(required_sheets - set(workbook.sheetnames))
    if missing:
        raise ValueError(f"Hidden Value formula contract requires support sheets: {missing!r}.")

    contract = load_json_strict(HIDDEN_VALUE_SIGNAL_CONTRACT)
    signals = sorted(
        (row for row in contract.get("signals") or [] if isinstance(row, dict)),
        key=lambda row: (int(row.get("priority") or 0), str(row.get("signal_id") or "")),
    )
    detail_specs = _hidden_value_detail_specs(signals)
    if len(signals) != 7 or len(detail_specs) != 91:
        raise ValueError("Hidden Value workbook surfaces require seven signals and 91 contract-derived detail rows.")

    recompute = workbook["Hidden_Value_Recompute"]
    for row, spec in enumerate(detail_specs, start=HIDDEN_VALUE_DETAIL_FIRST_ROW):
        _apply_hidden_value_detail_row(recompute, row, spec)
    for row, signal in enumerate(signals, start=HIDDEN_VALUE_CANDIDATE_FIRST_ROW):
        _apply_hidden_value_candidate_row(recompute, row, signal)

    audit = workbook["Hidden_Value_Audit"]
    for row in range(2, 9):
        _apply_hidden_value_audit_row(audit, row)
    _apply_hidden_value_defined_names(workbook)
    _apply_hidden_value_valuation_state_counts(workbook["Valuation"])


def _apply_hidden_value_valuation_state_counts(ws: Any) -> None:
    audit_states = "'Hidden_Value_Audit'!$R$2:$R$8"
    audit_keys = "'Hidden_Value_Audit'!$A$2:$A$8"
    formulas = {
        "R139": f'=IF(COUNTA({audit_keys})=0,"",COUNTIF({audit_states},"triggered"))',
        "R140": f'=IF(COUNTA({audit_keys})=0,"",COUNTIF({audit_states},"near_miss"))',
        "R141": f'=IF(COUNTA({audit_keys})=0,"",COUNTIF({audit_states},"not_triggered"))',
        "R142": f'=IF(COUNTA({audit_keys})=0,"",COUNTIF({audit_states},"insufficient_evidence"))',
        "R143": (
            f'=IF(COUNTA({audit_keys})=0,"",COUNTIF({audit_states},"unavailable")+'
            f'COUNTIF({audit_states},"invalid_input"))'
        ),
    }
    for coordinate, formula in formulas.items():
        _set_formula(ws[coordinate], formula, "0")


def _hidden_value_detail_specs(signals: list[dict[str, Any]]) -> list[dict[str, Any]]:
    specs: list[dict[str, Any]] = []
    for signal in signals:
        signal_id = str(signal.get("signal_id") or "")
        for metric_id in map(str, signal.get("required_metric_ids") or []):
            specs.append({"signal_id": signal_id, "record_type": "required_metric", "stage": "required", "item_id": metric_id})
        for stage in ("eligibility", "trigger", "near_miss"):
            for predicate in (signal.get(stage) or {}).get("predicates") or []:
                specs.append({"signal_id": signal_id, "record_type": "predicate", "stage": stage, "item_id": str(predicate.get("predicate_id") or "")})
        for component in signal.get("score_components") or []:
            specs.append({"signal_id": signal_id, "record_type": "score_component", "stage": "score", "item_id": str(component.get("component_id") or "")})
    return specs


def _apply_hidden_value_detail_row(ws: Any, row: int, spec: dict[str, Any]) -> None:
    record_type = str(spec["record_type"])
    module_eligible = _hidden_value_module_eligible_formula(row)
    metric_value = _hidden_value_metric_value_formula(f"$G{row}")
    metric_status = _hidden_value_metric_status_formula(f"$G{row}")
    value_formula = metric_value if record_type == "required_metric" else f'IF({module_eligible},{metric_value},"")'
    _set_formula(ws[f"X{row}"], f"={value_formula}", "0.0000000000")

    if record_type == "required_metric":
        _set_formula(ws[f"Y{row}"], f"={metric_status}", "General")
        _set_formula(ws[f"Z{row}"], f'=IF($G{row}="","",AND(ISNUMBER($X{row}),OR($Y{row}="source_backed",$Y{row}="formula_calculated",$Y{row}="derived_calculated")))', "General")
        for column in ("AA", "AB", "AC"):
            _set_formula(ws[f"{column}{row}"], '=""', "General")
        parity = "AND(" + ",".join((
            _hidden_value_equal_formula(f"$X{row}", f"$Q{row}"),
            _hidden_value_equal_formula(f"$Y{row}", f"$W{row}"),
            _hidden_value_equal_formula(f"$Z{row}", f"$S{row}"),
        )) + ")"
    elif record_type == "predicate":
        comparison = f'IF(NOT({module_eligible}),"",IF($I{row}<>"",{_hidden_value_metric_value_formula(f"$I{row}")},IF($J{row}="","",$J{row})))'
        _set_formula(ws[f"Y{row}"], f"={comparison}", "0.0000000000")
        _set_formula(ws[f"Z{row}"], f'=IF(OR($X{row}="",$Y{row}=""),"",IF($H{row}="gt",$X{row}>$Y{row},IF($H{row}="gte",$X{row}>=$Y{row},IF($H{row}="lt",$X{row}<$Y{row},IF($H{row}="lte",$X{row}<=$Y{row},IF($H{row}="eq",$X{row}=$Y{row},FALSE))))))', "General")
        for column in ("AA", "AB", "AC"):
            _set_formula(ws[f"{column}{row}"], '=""', "General")
        recomputed_reason = f'IF(NOT({module_eligible}),"module_or_profile_pack_disabled",IF($Z{row}="","predicate_input_unavailable",IF($Z{row},"passed","failed")))'
        parity = "AND(" + ",".join((
            _hidden_value_equal_formula(f"$X{row}", f"$Q{row}"),
            _hidden_value_equal_formula(f"$Y{row}", f"$R{row}"),
            _hidden_value_equal_formula(f"$Z{row}", f"$S{row}"),
            f"({recomputed_reason}=$W{row})",
        )) + ")"
    else:
        _set_formula(ws[f"Y{row}"], f'=IF({module_eligible},{metric_status},"unavailable")', "General")
        _set_formula(ws[f"Z{row}"], f'=IF(NOT({module_eligible}),FALSE,AND(ISNUMBER($X{row}),OR($Y{row}="source_backed",$Y{row}="formula_calculated",$Y{row}="derived_calculated")))', "General")
        _set_formula(ws[f"AA{row}"], f'=IF($Z{row},$L{row},0)', "0.0000000000")
        _set_formula(
            ws[f"AB{row}"],
            f'=IF(NOT($Z{row}),"",IF(OR(NOT(ISNUMBER($N{row})),NOT(ISNUMBER($O{row})),$O{row}<=0,NOT(ISNUMBER($P{row})),AND($M{row}<>"higher",$M{row}<>"lower")),"",MAX(0,MIN(1,IF($M{row}="lower",($N{row}-$X{row})/$O{row}+$P{row},($X{row}-$N{row})/$O{row}+$P{row})))))',
            "0.0000000000",
        )
        _set_formula(ws[f"AC{row}"], f'=IF($AB{row}="","",$AB{row}*$AA{row})', "0.0000000000")
        recomputed_status = f'IF($Z{row},"calculated","unavailable")'
        parity = "AND(" + ",".join((
            _hidden_value_equal_formula(f"$X{row}", f"$Q{row}"),
            _hidden_value_equal_formula(f"$AA{row}", f"$T{row}"),
            _hidden_value_equal_formula(f"$AB{row}", f"$U{row}"),
            _hidden_value_equal_formula(f"$AC{row}", f"$V{row}"),
            f"({recomputed_status}=$W{row})",
        )) + ")"
    _set_formula(ws[f"AD{row}"], f'=IF({parity},"PASS","FAIL")', "General")


def _apply_hidden_value_candidate_row(ws: Any, row: int, signal: dict[str, Any]) -> None:
    audit_row = row
    signal_id = str(signal.get("signal_id") or "")
    eligibility_mode = str((signal.get("eligibility") or {}).get("mode") or "all")
    trigger_mode = str((signal.get("trigger") or {}).get("mode") or "all")
    near_mode = str((signal.get("near_miss") or {}).get("mode") or "all")
    minimum = int((signal.get("near_miss") or {}).get("minimum_trigger_predicates") or 0)
    reweight = bool(signal.get("reweight_available_components"))
    formulas = {
        "AF": f'=IF(\'Hidden_Value_Audit\'!A{audit_row}="","",\'Hidden_Value_Audit\'!A{audit_row})',
        "AG": f'=IF($AF{row}="","","{signal_id}")',
        "AH": f'=IF($AF{row}="","","{eligibility_mode}")',
        "AI": f'=IF($AF{row}="","","{trigger_mode}")',
        "AJ": f'=IF($AF{row}="","","{near_mode}")',
        "AK": f'=IF($AF{row}="","",{minimum})',
        "AL": f'=IF($AF{row}="","",{str(reweight).upper()})',
        "AM": f'=IF($AF{row}="","",\'Hidden_Value_Audit\'!F{audit_row})',
        "AN": f'=IF($AF{row}="","",\'Hidden_Value_Audit\'!G{audit_row})',
        "AO": f'=IF(\'Hidden_Value_Audit\'!H{audit_row}="","",\'Hidden_Value_Audit\'!H{audit_row})',
        "AP": f'=IF($AF{row}="","",\'Hidden_Value_Audit\'!I{audit_row})',
    }
    for column, formula in formulas.items():
        _set_formula(ws[f"{column}{row}"], formula, "General")

    eligibility = _hidden_value_group_formula(row, "eligibility", mode_ref=f"$AH{row}")
    trigger = _hidden_value_group_formula(row, "trigger", mode_ref=f"$AI{row}")
    near_raw = _hidden_value_group_formula(row, "near_miss", mode_ref=f"$AJ{row}")
    trigger_true_count = _hidden_value_detail_count(row, record_type="predicate", stage="trigger", result=True)
    near = f'AND({near_raw},{trigger_true_count}>=$AK{row})'
    denominator = f'SUMIFS($AA$2:$AA$92,$B$2:$B$92,$AF{row},$D$2:$D$92,"score_component")'
    required_missing = _hidden_value_detail_count(row, record_type="score_component", required=True, result=False)
    weighted = f'SUMIFS($AC$2:$AC$92,$B$2:$B$92,$AF{row},$D$2:$D$92,"score_component")'
    raw_score = f'MAX(0,MIN(100,{weighted}*100/$AT{row}))'
    whole_score = f'INT({raw_score})'
    score = (
        f'IF(OR({required_missing}>0,$AT{row}<=0,AND(NOT($AL{row}),ABS($AT{row}-100)>1E-10)),"",'
        f'IF(ABS({raw_score}-{whole_score}-0.5)<=1E-12,'
        f'IF(MOD({whole_score},2)=0,{whole_score},{whole_score}+1),ROUND({raw_score},0)))'
    )
    required_invalid = _hidden_value_status_count(row, "invalid_input")
    required_insufficient = _hidden_value_status_count(row, "insufficient_evidence")
    required_unavailable = _hidden_value_detail_count(row, record_type="required_metric", result=False)
    trigger_available = _hidden_value_detail_count(row, record_type="predicate", stage="trigger", result="nonblank")
    module_eligible = f"'Hidden_Value_Audit'!$Q{audit_row}"
    state = (
        f'IF(NOT({module_eligible}),"unavailable",'
        f'IF({required_invalid}>0,"invalid_input",IF({required_insufficient}>0,"insufficient_evidence",'
        f'IF({required_unavailable}>0,"unavailable",IF(NOT($AQ{row}),"not_triggered",'
        f'IF({trigger_available}=0,"insufficient_evidence",IF(AND($AR{row},$AU{row}=""),"insufficient_evidence",'
        f'IF($AR{row},"triggered",IF($AS{row},"near_miss","not_triggered")))))))))'
    )
    detail_failures = f'COUNTIFS($B$2:$B$92,$AF{row},$AD$2:$AD$92,"<>PASS")'
    parity = "AND(" + ",".join((
        _hidden_value_equal_formula(f"$AV{row}", f"$AM{row}"),
        _hidden_value_equal_formula(f"$AW{row}", f"$AN{row}"),
        _hidden_value_equal_formula(f"$AU{row}", f"$AO{row}"),
        _hidden_value_equal_formula(f"$AT{row}", f"$AP{row}"),
        f"({detail_failures}=0)",
    )) + ")"
    _set_formula(ws[f"AQ{row}"], f'=IF($AF{row}="","",{eligibility})', "General")
    _set_formula(ws[f"AR{row}"], f'=IF($AF{row}="","",{trigger})', "General")
    _set_formula(ws[f"AS{row}"], f'=IF($AF{row}="","",{near})', "General")
    _set_formula(ws[f"AT{row}"], f'=IF($AF{row}="","",{denominator})', "0.0000000000")
    _set_formula(ws[f"AU{row}"], f'=IF($AF{row}="","",{score})', "0")
    _set_formula(ws[f"AV{row}"], f'=IF($AF{row}="","",{state})', "General")
    _set_formula(ws[f"AW{row}"], f'=IF($AF{row}="","",$AV{row}="triggered")', "General")
    _set_formula(ws[f"AX{row}"], f'=IF($AF{row}="","",IF({parity},"PASS","FAIL"))', "General")


def _apply_hidden_value_audit_row(ws: Any, row: int) -> None:
    match = f'MATCH($A{row},\'Hidden_Value_Recompute\'!$AF$2:$AF$8,0)'
    formulas = {
        "R": f'=IF($A{row}="","",IFERROR(INDEX(\'Hidden_Value_Recompute\'!$AV$2:$AV$8,{match}),""))',
        "S": f'=IF($A{row}="","",IFERROR(INDEX(\'Hidden_Value_Recompute\'!$AW$2:$AW$8,{match}),""))',
        "T": f'=IF($A{row}="","",IFERROR(INDEX(\'Hidden_Value_Recompute\'!$AU$2:$AU$8,{match}),""))',
        "U": f'=IF($A{row}="","",IFERROR(INDEX(\'Hidden_Value_Recompute\'!$AX$2:$AX$8,{match}),"FAIL"))',
        "V": f'=IF($A{row}="","",IF($U{row}="PASS","","hidden_value_recompute_mismatch"))',
    }
    for column, formula in formulas.items():
        _set_formula(ws[f"{column}{row}"], formula, "0" if column == "T" else "General")


def _hidden_value_module_eligible_formula(detail_row: int) -> str:
    return f'IFERROR(INDEX(\'Hidden_Value_Audit\'!$Q$2:$Q$8,MATCH($B{detail_row},\'Hidden_Value_Audit\'!$A$2:$A$8,0)),FALSE)'


def _hidden_value_metric_status_formula(key_ref: str) -> str:
    return f'IFERROR(INDEX(HV_Base_Status,MATCH({key_ref},HV_Base_MetricKey,0)),"")'


def _hidden_value_metric_value_formula(key_ref: str) -> str:
    status = _hidden_value_metric_status_formula(key_ref)
    value = f'IFERROR(INDEX(HV_Base_Value,MATCH({key_ref},HV_Base_MetricKey,0)),"")'
    return f'IF(OR({status}="source_backed",{status}="formula_calculated",{status}="derived_calculated"),{value},"")'


def _hidden_value_group_formula(candidate_row: int, stage: str, *, mode_ref: str) -> str:
    total = _hidden_value_detail_count(candidate_row, record_type="predicate", stage=stage)
    passed = _hidden_value_detail_count(candidate_row, record_type="predicate", stage=stage, result=True)
    return f'IF({mode_ref}="all",AND({total}>0,{passed}={total}),{passed}>0)'


def _hidden_value_detail_count(candidate_row: int, *, record_type: str, stage: str | None = None, required: bool | None = None, result: bool | str | None = None) -> str:
    criteria = [f'$B$2:$B$92,$AF{candidate_row}', f'$D$2:$D$92,"{record_type}"']
    if stage is not None:
        criteria.append(f'$E$2:$E$92,"{stage}"')
    if required is not None:
        criteria.append(f'$K$2:$K$92,{str(required).upper()}')
    if result == "nonblank":
        criteria.append('$Z$2:$Z$92,"<>"')
    elif isinstance(result, bool):
        criteria.append(f'$Z$2:$Z$92,{str(result).upper()}')
    return f"COUNTIFS({','.join(criteria)})"


def _hidden_value_status_count(candidate_row: int, status: str) -> str:
    return f'COUNTIFS($B$2:$B$92,$AF{candidate_row},$D$2:$D$92,"required_metric",$Y$2:$Y$92,"{status}")'


def _hidden_value_equal_formula(left: str, right: str) -> str:
    return (
        f'IF(OR({left}="",{right}=""),AND({left}="",{right}=""),'
        f'IF(AND(ISNUMBER({left}),ISNUMBER({right})),ABS({left}-{right})<=1E-10,{left}={right}))'
    )


def _remove_hidden_value_defined_names(workbook: Any) -> None:
    names = {
        "FCF_TTM_Pos_Years", "Pos_FCF_Ratio", "Interest_Coverage",
        "HV_Base_MetricKey", "HV_Base_Value", "HV_Base_Status",
        "HV_Recompute_CandidateKey", "HV_Recompute_RowParity", "HV_Recompute_CandidateParity",
        "HV_Flags_CandidateKey", "HV_Flags_Score", "HV_Flags_State",
    }
    for name in names:
        if name in workbook.defined_names:
            del workbook.defined_names[name]


def _apply_hidden_value_defined_names(workbook: Any) -> None:
    definitions = {
        "HV_Base_MetricKey": "'Hidden_Value_Base'!$A$2:$A$5001",
        "HV_Base_Value": "'Hidden_Value_Base'!$C$2:$C$5001",
        "HV_Base_Status": "'Hidden_Value_Base'!$I$2:$I$5001",
        "HV_Recompute_CandidateKey": "'Hidden_Value_Recompute'!$AF$2:$AF$8",
        "HV_Recompute_RowParity": "'Hidden_Value_Recompute'!$AD$2:$AD$92",
        "HV_Recompute_CandidateParity": "'Hidden_Value_Recompute'!$AX$2:$AX$8",
        "HV_Flags_CandidateKey": "'Hidden_Value_Flags'!$B$2:$B$8",
        "HV_Flags_Score": "'Hidden_Value_Flags'!$E$2:$E$8",
        "HV_Flags_State": "'Hidden_Value_Flags'!$G$2:$G$8",
    }
    for name, attr_text in definitions.items():
        workbook.defined_names.add(DefinedName(name, attr_text=attr_text))


def _set_formula(cell: Any, formula: str, number_format: str) -> None:
    cell.value = formula
    cell.number_format = number_format
    protection = copy(cell.protection)
    protection.locked = True
    cell.protection = protection


def _number_format_for_row(row: int) -> str:
    for contract in FORMULA_ROWS:
        if contract.row == row:
            return contract.number_format
    return "General"


def _history_range(column: str) -> str:
    return f"History_Q!${column}${CALCULATION_HISTORY_FIRST_ROW}:${column}${CALCULATION_HISTORY_LAST_ROW}"


def _history_ordinal(period_cell: str, metric: str) -> str:
    return (
        f'SUMIFS({_history_range(CALCULATION_HISTORY_COLUMNS["period_ordinal"])},'
        f'{_history_range(CALCULATION_HISTORY_COLUMNS["period"])},{period_cell},'
        f'{_history_range(CALCULATION_HISTORY_COLUMNS["metric"])},"{metric}")'
    )


def _history_end_ordinal(period_cell: str, metric: str, offset: int = 0) -> str:
    base = _history_ordinal(period_cell, metric)
    if offset == 0:
        return f"({base})"
    sign = "+" if offset > 0 else ""
    return f"({base}{sign}{offset})"


def _history_window_terms(period_cell: str, metric: str, unit: str, *, end_offset: int = 0) -> tuple[str, str, str, str]:
    end = _history_end_ordinal(period_cell, metric, end_offset)
    start = f"({end}-3)"
    metric_range = _history_range(CALCULATION_HISTORY_COLUMNS["metric"])
    ordinal_range = _history_range(CALCULATION_HISTORY_COLUMNS["period_ordinal"])
    unit_range = _history_range(CALCULATION_HISTORY_COLUMNS["unit"])
    status_range = _history_range(CALCULATION_HISTORY_COLUMNS["status"])
    value_range = _history_range(CALCULATION_HISTORY_COLUMNS["value"])
    criteria = (
        f'{metric_range},"{metric}",{ordinal_range},">="&{start},'
        f'{ordinal_range},"<="&{end},{unit_range},"{unit}",{status_range},"populated"'
    )
    count = f"COUNTIFS({criteria})"
    total = f"SUMIFS({value_range},{criteria})"
    minimum = f"MINIFS({ordinal_range},{metric_range},\"{metric}\",{ordinal_range},\">=\"&{start},{ordinal_range},\"<=\"&{end},{unit_range},\"{unit}\",{status_range},\"populated\")"
    maximum = f"MAXIFS({ordinal_range},{metric_range},\"{metric}\",{ordinal_range},\">=\"&{start},{ordinal_range},\"<=\"&{end},{unit_range},\"{unit}\",{status_range},\"populated\")"
    return count, total, minimum, maximum


def _history_point_terms(period_cell: str, metric: str, unit: str, *, offset: int = 0) -> tuple[str, str]:
    ordinal = _history_end_ordinal(period_cell, metric, offset)
    metric_range = _history_range(CALCULATION_HISTORY_COLUMNS["metric"])
    ordinal_range = _history_range(CALCULATION_HISTORY_COLUMNS["period_ordinal"])
    unit_range = _history_range(CALCULATION_HISTORY_COLUMNS["unit"])
    status_range = _history_range(CALCULATION_HISTORY_COLUMNS["status"])
    value_range = _history_range(CALCULATION_HISTORY_COLUMNS["value"])
    criteria = f'{metric_range},"{metric}",{ordinal_range},{ordinal},{unit_range},"{unit}",{status_range},"populated"'
    return f"COUNTIFS({criteria})", f"SUMIFS({value_range},{criteria})"


def _history_ttm_sum(period_cell: str, metric: str, unit: str, *, end_offset: int = 0) -> str:
    count, total, minimum, maximum = _history_window_terms(period_cell, metric, unit, end_offset=end_offset)
    return f'=IF(OR({period_cell}="",{count}<>4,{maximum}-{minimum}<>3),"",{total})'


def _history_ttm_ratio(period_cell: str, numerator_metric: str, denominator_metric: str, unit: str) -> str:
    n_count, n_total, n_min, n_max = _history_window_terms(period_cell, numerator_metric, unit)
    d_count, d_total, d_min, d_max = _history_window_terms(period_cell, denominator_metric, unit)
    return f'=IF(OR({period_cell}="",{n_count}<>4,{d_count}<>4,{n_max}-{n_min}<>3,{d_max}-{d_min}<>3,{d_total}=0),"",{n_total}/{d_total})'


def _history_yoy_ratio(period_cell: str, metric: str, unit: str) -> str:
    current_count, current = _history_point_terms(period_cell, metric, unit)
    prior_count, prior = _history_point_terms(period_cell, metric, unit, offset=-4)
    return f'=IF(OR({period_cell}="",{current_count}<>1,{prior_count}<>1,{prior}=0),"",{current}/{prior}-1)'


def _history_point_difference(period_cell: str, metric: str, unit: str, *, offset: int) -> str:
    current_count, current = _history_point_terms(period_cell, metric, unit)
    prior_count, prior = _history_point_terms(period_cell, metric, unit, offset=offset)
    return f'=IF(OR({period_cell}="",{current_count}<>1,{prior_count}<>1),"",{current}-{prior})'


def _history_ttm_difference(period_cell: str, left_metric: str, right_metric: str, unit: str, *, end_offset: int = 0) -> str:
    l_count, left, l_min, l_max = _history_window_terms(period_cell, left_metric, unit, end_offset=end_offset)
    r_count, right, r_min, r_max = _history_window_terms(period_cell, right_metric, unit, end_offset=end_offset)
    return f'=IF(OR({period_cell}="",{l_count}<>4,{r_count}<>4,{l_max}-{l_min}<>3,{r_max}-{r_min}<>3),"",{left}-{right})'


def _history_yoy_difference_of_differences(period_cell: str, left_metric: str, right_metric: str, unit: str) -> str:
    lc_count, left_current = _history_point_terms(period_cell, left_metric, unit)
    rc_count, right_current = _history_point_terms(period_cell, right_metric, unit)
    lp_count, left_prior = _history_point_terms(period_cell, left_metric, unit, offset=-4)
    rp_count, right_prior = _history_point_terms(period_cell, right_metric, unit, offset=-4)
    return f'=IF(OR({period_cell}="",{lc_count}<>1,{rc_count}<>1,{lp_count}<>1,{rp_count}<>1),"",({left_current}-{right_current})-({left_prior}-{right_prior}))'


def _history_ttm_yoy_difference_of_differences(period_cell: str, left_metric: str, right_metric: str, unit: str) -> str:
    current_left = _history_window_terms(period_cell, left_metric, unit)
    current_right = _history_window_terms(period_cell, right_metric, unit)
    prior_left = _history_window_terms(period_cell, left_metric, unit, end_offset=-4)
    prior_right = _history_window_terms(period_cell, right_metric, unit, end_offset=-4)
    checks = []
    for count, _total, minimum, maximum in (current_left, current_right, prior_left, prior_right):
        checks.extend((f"{count}<>4", f"{maximum}-{minimum}<>3"))
    current = f"({current_left[1]}-{current_right[1]})"
    prior = f"({prior_left[1]}-{prior_right[1]})"
    return f'=IF(OR({period_cell}="",{",".join(checks)}),"",{current}-{prior})'


def _ttm_sum(start_col: str, end_col: str, row: int) -> str:
    if not start_col:
        return '=""'
    range_ref = f"{start_col}{row}:{end_col}{row}"
    return f'=IF(COUNT({range_ref})<4,"",SUM({range_ref}))'


def _ttm_ratio(start_col: str, end_col: str, numerator_row: int, denominator_row: int) -> str:
    if not start_col:
        return '=""'
    numerator = f"{start_col}{numerator_row}:{end_col}{numerator_row}"
    denominator = f"{start_col}{denominator_row}:{end_col}{denominator_row}"
    return f'=IF(OR(COUNT({numerator})<4,COUNT({denominator})<4,SUM({denominator})=0),"",SUM({numerator})/SUM({denominator}))'


def _ratio(numerator: str, denominator: str) -> str:
    return f'=IF(OR({numerator}="",{denominator}="",{denominator}=0),"",{numerator}/{denominator})'


def _difference(left: str, right: str) -> str:
    return f'=IF(OR({left}="",{right}=""),"",{left}-{right})'


def _yoy_ratio(prior_col: str, current_col: str, row: int) -> str:
    if not prior_col:
        return '=""'
    return f'=IF(OR({prior_col}{row}="",{current_col}{row}="",{prior_col}{row}=0),"",{current_col}{row}/{prior_col}{row}-1)'


def _yoy_difference(prior_col: str, current_col: str, row: int) -> str:
    if not prior_col:
        return '=""'
    return _difference(f"{current_col}{row}", f"{prior_col}{row}")


def _qoq_difference(column: int, row: int) -> str:
    if column <= FIRST_QUARTER_COLUMN:
        return '=""'
    current = get_column_letter(column)
    prior = get_column_letter(column - 1)
    return _difference(f"{current}{row}", f"{prior}{row}")


def _bs_sales_yoy_formula(column: str) -> str:
    lookup = f"MATCH({column}$7,'Valuation'!$B$6:$M$6,0)"
    current = f"INDEX('Valuation'!$B$9:$M$9,1,{lookup})"
    prior = f"INDEX('Valuation'!$B$9:$M$9,1,{lookup}-4)"
    return f'=IFERROR(IF(OR({current}="",{prior}="",{prior}=0),"",{current}/{prior}-1),"")'
