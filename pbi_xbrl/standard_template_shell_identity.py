"""Deterministic identity contract for the frozen standard workbook shell."""
from __future__ import annotations

import hashlib
import json
import math
import re
import zipfile
import xml.etree.ElementTree as ET
from collections.abc import Iterator, Mapping as MappingABC
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.utils import range_boundaries

from pbi_xbrl.json_schema_validation import load_json_strict, validate_json_schema


SHELL_SEMANTIC_CONTRACT_VERSION = "1.6.0"
SHEET_VIEW_IDENTITY_CONTRACT_VERSION = "1.0.0"
BINDING_PLANNER_CONTRACT_VERSION = "1.3.0"
ROOT = Path(__file__).resolve().parents[1]
MANIFEST_SCHEMA_PATH = ROOT / "docs" / "standard_template_shell_manifest.schema.json"
BINDING_SCHEMA_PATH = ROOT / "docs" / "workbook_binding_map.schema.json"
BINDING_PLAN_SCHEMA_PATH = ROOT / "docs" / "new_ticker_binding_plan.schema.json"
IDENTITY_FIELDS = (
    "workbook_sha256",
    "manifest_contract_signature",
    "sheet_order_visibility_signature",
    "sheet_view_signature",
    "merge_signature",
    "defined_name_signature",
    "writable_target_signature",
    "binding_contract_signature",
    "formula_static_zone_signature",
)
_FIXED_ZIP_TIMESTAMP = (2000, 1, 1, 0, 0, 0)
_FIXED_CORE_TIMESTAMP = "2000-01-01T00:00:00Z"
_CORE_NAMESPACES = {
    "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
    "dc": "http://purl.org/dc/elements/1.1/",
    "dcterms": "http://purl.org/dc/terms/",
    "dcmitype": "http://purl.org/dc/dcmitype/",
    "xsi": "http://www.w3.org/2001/XMLSchema-instance",
}
_SHEET_VIEW_BOOLEAN_DEFAULTS = {
    "windowProtection": False,
    "showFormulas": False,
    "showGridLines": True,
    "showRowColHeaders": True,
    "showZeros": True,
    "rightToLeft": False,
    "tabSelected": False,
    "showRuler": True,
    "showOutlineSymbols": True,
    "defaultGridColor": True,
    "showWhiteSpace": True,
    "zoomToFit": False,
}
SHEET_VIEW_CONTRACT_OWNED_PROPERTIES = (
    *_SHEET_VIEW_BOOLEAN_DEFAULTS,
    "view",
    "workbookViewId",
    "colorId",
    "zoomScale",
    "zoomScaleNormal",
    "zoomScaleSheetLayoutView",
    "zoomScalePageLayoutView",
    "freezePanes",
    "pane.xSplit",
    "pane.ySplit",
    "pane.topLeftCell",
    "pane.activePane",
    "pane.state",
)
SHEET_VIEW_CONTRACT_IGNORED_PROPERTIES = (
    "sheetView.topLeftCell",
    "selection.activeCell",
    "selection.activeCellId",
    "selection.sqref",
    "selection.pane",
    "workbook.activeTab",
)
_VERIFICATION_AUTHORITY = object()


@dataclass(frozen=True)
class VerifiedShellIdentity(MappingABC[str, Any]):
    """Authenticated result produced only by real shell verification."""

    status: str
    issues: tuple[dict[str, str], ...]
    expected: dict[str, Any]
    actual: dict[str, Any]
    _authority: object = field(repr=False, compare=False)

    def __getitem__(self, key: str) -> Any:
        return self.to_dict()[key]

    def __iter__(self) -> Iterator[str]:
        return iter(("status", "issues", "expected", "actual"))

    def __len__(self) -> int:
        return 4

    def to_dict(self) -> dict[str, Any]:
        return {
            "status": self.status,
            "issues": [dict(issue) for issue in self.issues],
            "expected": dict(self.expected),
            "actual": dict(self.actual),
        }


def is_verified_shell_identity(value: Any) -> bool:
    return isinstance(value, VerifiedShellIdentity) and value._authority is _VERIFICATION_AUTHORITY


def _verified_result(
    *,
    status: str,
    issues: Sequence[Mapping[str, Any]],
    expected: Mapping[str, Any],
    actual: Mapping[str, Any],
) -> VerifiedShellIdentity:
    return VerifiedShellIdentity(
        status=status,
        issues=tuple(
            {
                "rule_id": str(issue.get("rule_id") or "shell_identity_failure"),
                "message": str(issue.get("message") or "Shell identity verification failed."),
            }
            for issue in issues
        ),
        expected=dict(expected),
        actual=dict(actual),
        _authority=_VERIFICATION_AUTHORITY,
    )


def normalize_xlsx_package(path: Path | str) -> Path:
    """Rewrite ZIP metadata deterministically without changing workbook XML."""

    workbook_path = Path(path)
    entries: list[tuple[str, bytes, int]] = []
    with zipfile.ZipFile(workbook_path, "r") as source:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename == "docProps/core.xml":
                data = _normalize_core_properties(data)
            entries.append((info.filename, data, info.external_attr))
    with NamedTemporaryFile(prefix=workbook_path.stem + "-", suffix=".xlsx", dir=workbook_path.parent, delete=False) as handle:
        temporary = Path(handle.name)
    try:
        with zipfile.ZipFile(temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as target:
            for name, data, external_attr in sorted(entries, key=lambda row: row[0]):
                info = zipfile.ZipInfo(name, date_time=_FIXED_ZIP_TIMESTAMP)
                info.compress_type = zipfile.ZIP_DEFLATED
                info.create_system = 0
                info.external_attr = external_attr
                target.writestr(info, data)
        temporary.replace(workbook_path)
    finally:
        if temporary.exists():
            temporary.unlink()
    return workbook_path


def _normalize_core_properties(data: bytes) -> bytes:
    for prefix, uri in _CORE_NAMESPACES.items():
        ET.register_namespace(prefix, uri)
    root = ET.fromstring(data)
    for local_name in ("created", "modified"):
        node = root.find(f"{{{_CORE_NAMESPACES['dcterms']}}}{local_name}")
        if node is not None:
            node.text = _FIXED_CORE_TIMESTAMP
    return ET.tostring(root, encoding="utf-8", xml_declaration=False)


def _normalized_number(value: Any) -> int | float | None:
    if value is None:
        return None
    number = float(value)
    return int(number) if number.is_integer() else number


def _normalized_sheet_view(view: Any) -> dict[str, Any]:
    pane = getattr(view, "pane", None)
    zoom_scale = getattr(view, "zoomScale", None)
    pane_contract = None
    if pane is not None:
        pane_contract = {
            "xSplit": _normalized_number(getattr(pane, "xSplit", None)),
            "ySplit": _normalized_number(getattr(pane, "ySplit", None)),
            "topLeftCell": str(getattr(pane, "topLeftCell", "") or ""),
            "activePane": str(getattr(pane, "activePane", "") or ""),
            "state": str(getattr(pane, "state", "") or ""),
        }
    contract = {
        name: default if getattr(view, name, None) is None else bool(getattr(view, name))
        for name, default in _SHEET_VIEW_BOOLEAN_DEFAULTS.items()
    }
    contract.update(
        {
            "view": str(getattr(view, "view", None) or "normal"),
            "workbookViewId": int(getattr(view, "workbookViewId", None) or 0),
            "colorId": (
                int(getattr(view, "colorId"))
                if getattr(view, "colorId", None) is not None
                else None
            ),
            "zoomScale": int(zoom_scale) if zoom_scale is not None else 100,
            "zoomScaleNormal": (
                int(getattr(view, "zoomScaleNormal"))
                if getattr(view, "zoomScaleNormal", None) is not None
                else None
            ),
            "zoomScaleSheetLayoutView": (
                int(getattr(view, "zoomScaleSheetLayoutView"))
                if getattr(view, "zoomScaleSheetLayoutView", None) is not None
                else None
            ),
            "zoomScalePageLayoutView": (
                int(getattr(view, "zoomScalePageLayoutView"))
                if getattr(view, "zoomScalePageLayoutView", None) is not None
                else None
            ),
            "pane": pane_contract,
        }
    )
    return contract


def _sheet_view_contract_payload(
    wb: Any,
    *,
    sheet_name_map: Mapping[str, str] | None = None,
) -> dict[str, Any]:
    name_map = sheet_name_map or {ws.title: ws.title for ws in wb.worksheets}
    worksheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        views = list(getattr(getattr(ws, "views", None), "sheetView", ()) or ())
        worksheets.append(
            {
                "sheet": str(name_map.get(ws.title, ws.title)),
                "freezePanes": str(ws.freeze_panes or ""),
                "views": [_normalized_sheet_view(view) for view in views],
            }
        )
    return {
        "contract_version": SHEET_VIEW_IDENTITY_CONTRACT_VERSION,
        "worksheets": worksheets,
    }


def compute_shell_identity(
    workbook_path: Path | str,
    *,
    manifest: Mapping[str, Any],
    binding_payload: Mapping[str, Any] | Sequence[Mapping[str, Any]],
    semantic_contract_version: str | None = None,
) -> dict[str, Any]:
    path = Path(workbook_path)
    version = semantic_contract_version or str(manifest.get("semantic_contract_version") or SHELL_SEMANTIC_CONTRACT_VERSION)
    bindings = list(binding_payload.get("bindings") or []) if isinstance(binding_payload, Mapping) else list(binding_payload)
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        sheets = [{"sheet": ws.title, "state": ws.sheet_state} for ws in wb.worksheets]
        sheet_views = _sheet_view_contract_payload(wb)
        merges = sorted(f"{ws.title}!{merged}" for ws in wb.worksheets for merged in ws.merged_cells.ranges)
        defined_names = sorted(
            (
                {
                "name": str(name),
                "attr_text": str(getattr(wb.defined_names[name], "attr_text", "") or ""),
                "type": str(getattr(wb.defined_names[name], "type", "") or ""),
                }
                for name in wb.defined_names
            ),
            key=lambda row: (row["name"], row["attr_text"], row["type"]),
        )
        writable_zones = _writable_zone_map(manifest)
        static_cells: list[dict[str, Any]] = []
        formula_cells: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            zones = writable_zones.get(ws.title, ())
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if value in (None, "") or _cell_in_zones(cell.column, cell.row, zones):
                        continue
                    record = {"sheet": ws.title, "cell": cell.coordinate, "value": str(value)}
                    if isinstance(value, str) and value.startswith("="):
                        formula_cells.append(record)
                    elif isinstance(value, (str, int, float, bool)):
                        static_cells.append(record)
    finally:
        wb.close()

    target_contract = _writable_target_contract(manifest, bindings)
    binding_contract = _executable_binding_contract(bindings)
    formula_static_payload = {
        "formula_cells": sorted(formula_cells, key=lambda row: (row["sheet"], row["cell"])),
        "static_cells": sorted(static_cells, key=lambda row: (row["sheet"], row["cell"])),
        "non_writable_zones": _non_writable_contract(manifest),
    }
    return {
        "semantic_contract_version": version,
        "workbook_sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "manifest_contract_signature": compute_manifest_contract_signature(manifest),
        "sheet_order_visibility_signature": _signature(sheets),
        "sheet_view_signature": _signature(sheet_views),
        "merge_signature": _signature(merges),
        "defined_name_signature": _signature(defined_names),
        "writable_target_signature": _signature(target_contract),
        "binding_contract_signature": compute_binding_contract_signature(binding_payload),
        "formula_static_zone_signature": _signature(formula_static_payload),
        "counts": {
            "sheets": len(sheets),
            "sheet_views": sum(len(row["views"]) for row in sheet_views["worksheets"]),
            "merges": len(merges),
            "defined_names": len(defined_names),
            "writable_targets": len(target_contract),
            "binding_contracts": len(binding_contract),
            "formula_cells": len(formula_cells),
            "static_cells": len(static_cells),
        },
    }


def compute_binding_contract_signature(
    binding_payload: Mapping[str, Any] | Sequence[Mapping[str, Any]],
) -> str:
    """Return the canonical signature for the complete binding-map contract.

    A binding list is not sufficient identity: planner versions, top-level
    defaults, and future policies can change execution without changing an
    individual binding. Mapping payloads are therefore signed in full. Sequence
    input remains supported only as an explicitly versioned legacy contract.
    """

    return _signature(_binding_document_contract(binding_payload))


def compute_manifest_contract_signature(manifest: Mapping[str, Any]) -> str:
    """Sign every manifest semantic without creating a circular hash contract."""

    payload = {
        str(key): _contract_value(value)
        for key, value in manifest.items()
        if str(key) != "shell_identity"
    }
    return _signature(payload)


def validate_verified_shell_token(
    report: VerifiedShellIdentity | None,
    *,
    manifest: Mapping[str, Any],
    binding_payload: Mapping[str, Any] | Sequence[Mapping[str, Any]],
) -> list[dict[str, str]]:
    """Validate that a token was issued for this complete manifest and binding contract."""

    if not is_verified_shell_identity(report):
        return [
            {
                "rule_id": "shell_identity_not_verified",
                "message": "A real VerifiedShellIdentity token is required.",
            }
        ]
    issues: list[dict[str, str]] = [dict(issue) for issue in report.issues]
    if report.status != "PASS":
        issues.append(
            {
                "rule_id": "shell_identity_token_failed",
                "message": "The supplied shell identity token did not pass verification.",
            }
        )
    expected = manifest.get("shell_identity") if isinstance(manifest.get("shell_identity"), Mapping) else {}
    current_version = str(manifest.get("semantic_contract_version") or "")
    current_contract = {
        "semantic_contract_version": current_version,
        **{field: str(expected.get(field) or "") for field in IDENTITY_FIELDS},
    }
    token_expected = {
        "semantic_contract_version": str(report.expected.get("semantic_contract_version") or ""),
        **{field: str(report.expected.get(field) or "") for field in IDENTITY_FIELDS},
    }
    token_actual = {
        "semantic_contract_version": str(report.actual.get("semantic_contract_version") or ""),
        **{field: str(report.actual.get(field) or "") for field in IDENTITY_FIELDS},
    }
    if current_contract != token_expected or current_contract != token_actual:
        issues.append(
            {
                "rule_id": "shell_identity_token_manifest_mismatch",
                "message": "The verified token was issued for a different manifest or shell identity contract.",
            }
        )
    if current_version != SHELL_SEMANTIC_CONTRACT_VERSION:
        issues.append(
            {
                "rule_id": "shell_semantic_contract_version_unsupported",
                "message": f"Unsupported semantic contract version {current_version!r}.",
            }
        )
    current_manifest_signature = compute_manifest_contract_signature(manifest)
    if current_manifest_signature != str(expected.get("manifest_contract_signature") or ""):
        issues.append(
            {
                "rule_id": "shell_manifest_contract_drift",
                "message": "The manifest semantics differ from its approved manifest-contract signature.",
            }
        )
    current_binding_signature = compute_binding_contract_signature(binding_payload)
    if current_binding_signature != str(expected.get("binding_contract_signature") or ""):
        issues.append(
            {
                "rule_id": "shell_binding_contract_token_mismatch",
                "message": "The executable binding contract differs from the contract approved by the token.",
            }
        )
    deduped: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for issue in issues:
        key = (str(issue.get("rule_id") or ""), str(issue.get("message") or ""))
        if key not in seen:
            seen.add(key)
            deduped.append(issue)
    return deduped


def verify_shell_identity(
    workbook_path: Path | str,
    *,
    manifest: Mapping[str, Any],
    binding_payload: Mapping[str, Any] | Sequence[Mapping[str, Any]],
) -> VerifiedShellIdentity:
    expected = manifest.get("shell_identity") if isinstance(manifest.get("shell_identity"), Mapping) else {}
    issues: list[dict[str, str]] = []
    issues.extend(_contract_schema_issues(manifest, MANIFEST_SCHEMA_PATH, rule_prefix="shell_manifest_schema"))
    binding_document: Mapping[str, Any]
    if isinstance(binding_payload, Mapping):
        binding_document = binding_payload
    else:
        binding_document = {
            "version": "0.3.0",
            "binding_planner_contract_version": "1.0.0",
            "bindings": list(binding_payload),
        }
    issues.extend(_contract_schema_issues(binding_document, BINDING_SCHEMA_PATH, rule_prefix="shell_binding_schema"))
    if not expected:
        return _verified_result(
            status="FAIL",
            issues=[{"rule_id": "shell_identity_missing", "message": "Manifest has no shell_identity contract."}],
            expected={},
            actual={},
        )
    manifest_version = str(manifest.get("semantic_contract_version") or "")
    expected_version = str(expected.get("semantic_contract_version") or "")
    if manifest_version != SHELL_SEMANTIC_CONTRACT_VERSION:
        issues.append(
            {
                "rule_id": "shell_semantic_contract_version_unsupported",
                "message": (
                    f"supported={SHELL_SEMANTIC_CONTRACT_VERSION!r} "
                    f"manifest={manifest_version!r}"
                ),
            }
        )
    if not manifest_version or manifest_version != expected_version:
        issues.append(
            {
                "rule_id": "shell_semantic_contract_version_mismatch",
                "message": f"manifest={manifest_version!r} identity={expected_version!r}",
            }
        )
    try:
        actual = compute_shell_identity(
            workbook_path,
            manifest=manifest,
            binding_payload=binding_payload,
            semantic_contract_version=manifest_version or expected_version,
        )
    except Exception as exc:
        return _verified_result(
            status="FAIL",
            issues=[*issues, {"rule_id": "shell_identity_unavailable", "message": str(exc)}],
            expected=expected,
            actual={},
        )
    for field in IDENTITY_FIELDS:
        if str(actual.get(field) or "") != str(expected.get(field) or ""):
            issues.append(
                {
                    "rule_id": _identity_rule_id(field),
                    "message": f"{field} differs: expected={expected.get(field)!r} actual={actual.get(field)!r}",
                }
            )
    return _verified_result(
        status="PASS" if not issues else "FAIL",
        issues=issues,
        expected=expected,
        actual=actual,
    )


def verify_post_fill_structural_identity(
    filled_workbook: Any,
    *,
    approved_shell_path: Path | str,
    manifest: Mapping[str, Any],
    binding_payload: Mapping[str, Any] | Sequence[Mapping[str, Any]],
    approved_plan: Any = None,
    normalized_package: Mapping[str, Any] | None = None,
    module_payload: Mapping[str, Any] | None = None,
    style_contract: Mapping[str, Any] | None = None,
    approved_style_plan: Any = None,
) -> dict[str, Any]:
    """Verify a filled workbook against an exact, approved source shell.

    Writable cell values and the resolved ticker investment-case sheet name may
    differ. Every protected value/formula, style, merge, defined name, layout,
    validation and table signature must remain structurally identical.
    """

    source_token = verify_shell_identity(
        approved_shell_path,
        manifest=manifest,
        binding_payload=binding_payload,
    )
    issues: list[dict[str, str]] = [dict(issue) for issue in source_token.issues]
    if source_token.status != "PASS":
        issues.append(
            {
                "rule_id": "post_fill_source_shell_not_approved",
                "message": "The source shell does not match the approved manifest identity.",
            }
        )
        return {"status": "FAIL", "issues": issues, "source_identity": source_token.to_dict()}

    reproduced_plan = None
    reproduced_style_plan = None
    plan_participates = normalized_package is not None or approved_plan is not None
    if plan_participates:
        if normalized_package is None:
            issues.append(
                {
                    "rule_id": "post_fill_normalized_package_required",
                    "message": "Post-fill validation requires the normalized package for independent plan reproduction.",
                }
            )
        else:
            style_mode = module_payload is not None or style_contract is not None or (
                isinstance(binding_payload, Mapping) and binding_payload.get("module_profile_id")
            )
            if style_mode:
                from pbi_xbrl.new_ticker_binding_planner import BindingPlanReproductionError
                from pbi_xbrl.new_ticker_style_planner import StylePlanningError, reproduce_style_plan

                try:
                    reproduced_plan, reproduced_style_plan = reproduce_style_plan(
                        normalized_package,
                        manifest=manifest,
                        binding_payload=binding_payload,
                        shell_path=approved_shell_path,
                        module_payload=module_payload,
                        style_contract=style_contract,
                        expected_binding_plan=approved_plan,
                        expected_style_plan=approved_style_plan,
                    )
                except (BindingPlanReproductionError, StylePlanningError) as exc:
                    issues.append(
                        {
                            "rule_id": "post_fill_binding_plan_reproduction_failed",
                            "message": str(exc),
                        }
                    )
            else:
                from pbi_xbrl.new_ticker_binding_planner import (
                    BindingPlanReproductionError,
                    reproduce_binding_plan,
                )

                try:
                    reproduced_plan = reproduce_binding_plan(
                        normalized_package,
                        manifest=manifest,
                        binding_payload=binding_payload,
                        shell_path=approved_shell_path,
                        expected_plan=approved_plan,
                    )
                except BindingPlanReproductionError as exc:
                    issues.append(
                        {
                            "rule_id": "post_fill_binding_plan_reproduction_failed",
                            "message": str(exc),
                        }
                    )

    source_wb = load_workbook(Path(approved_shell_path), read_only=False, data_only=False)
    owns_filled = isinstance(filled_workbook, (str, Path))
    target_wb = load_workbook(Path(filled_workbook), read_only=False, data_only=False) if owns_filled else filled_workbook
    try:
        if reproduced_style_plan is not None:
            from pbi_xbrl.new_ticker_style_application import apply_style_plan

            apply_style_plan(source_wb, reproduced_style_plan)
        bindings = list(binding_payload.get("bindings") or []) if isinstance(binding_payload, Mapping) else list(binding_payload)
        allowed_cells = _exact_writable_cells(bindings)
        source_payload = _post_fill_structural_payload(source_wb, allowed_cells=allowed_cells)
        target_payload = _post_fill_structural_payload(target_wb, allowed_cells=allowed_cells)
        source_values = _writable_value_map(source_wb, allowed_cells=allowed_cells)
        target_values = _writable_value_map(target_wb, allowed_cells=allowed_cells)
    finally:
        source_wb.close()
        if owns_filled:
            target_wb.close()

    source_signatures = {key: _signature(value) for key, value in source_payload.items()}
    target_signatures = {key: _signature(value) for key, value in target_payload.items()}
    rule_ids = {
        "sheets": "post_fill_sheet_order_visibility_drift",
        "sheet_views": "post_fill_sheet_view_drift",
        "merges": "post_fill_merge_drift",
        "defined_names": "post_fill_defined_name_drift",
        "layout": "post_fill_layout_drift",
        "cell_structure": "post_fill_protected_cell_drift",
        "data_validations": "post_fill_data_validation_drift",
        "conditional_formatting": "post_fill_conditional_formatting_drift",
        "tables": "post_fill_table_drift",
    }
    for section, source_signature in source_signatures.items():
        if target_signatures.get(section) != source_signature:
            issues.append(
                {
                    "rule_id": rule_ids[section],
                    "message": f"Filled workbook structural section {section!r} differs from the approved source shell.",
                }
            )
    changed_values = {
        key: target_values.get(key)
        for key in sorted(set(source_values) | set(target_values))
        if _contract_value(source_values.get(key)) != _contract_value(target_values.get(key))
    }
    plan_participates = plan_participates or bool(changed_values)
    if plan_participates:
        if normalized_package is None:
            if not any(issue.get("rule_id") == "post_fill_normalized_package_required" for issue in issues):
                issues.append(
                    {
                        "rule_id": "post_fill_normalized_package_required",
                        "message": "Post-fill validation requires the normalized package for independent plan reproduction.",
                    }
                )
        elif reproduced_plan is not None:
            plan_payload = reproduced_plan.to_dict()
            planned_values = _planned_value_map(plan_payload)
            allowed_keys = {f"{sheet}!{cell}" for sheet, cells in allowed_cells.items() for cell in cells}
            unauthorized_plan_targets = sorted(set(planned_values) - allowed_keys)
            if unauthorized_plan_targets:
                issues.append(
                    {
                        "rule_id": "post_fill_plan_target_not_owned",
                        "message": f"Plan contains {len(unauthorized_plan_targets)} target(s) outside exact active binding ownership.",
                    }
                )
            unauthorized_changes = sorted(set(changed_values) - set(planned_values))
            if unauthorized_changes:
                issues.append(
                    {
                        "rule_id": "post_fill_unplanned_value_change",
                        "message": f"Filled workbook contains {len(unauthorized_changes)} changed writable cell(s) absent from the approved plan.",
                    }
                )
            mismatches = [
                key
                for key, value in planned_values.items()
                if not _planned_cell_values_equal(target_values.get(key), value)
            ]
            if mismatches:
                issues.append(
                    {
                        "rule_id": "post_fill_planned_value_mismatch",
                        "message": f"Filled workbook does not match {len(mismatches)} approved planned value(s).",
                    }
                )
    return {
        "status": "PASS" if not issues else "FAIL",
        "issues": issues,
        "source_identity": source_token.to_dict(),
        "source_structural_signatures": source_signatures,
        "filled_structural_signatures": target_signatures,
        "changed_writable_cell_count": len(changed_values),
        "reproduced_style_action_count": (
            len(reproduced_style_plan.actions) if reproduced_style_plan is not None else 0
        ),
    }


def _writable_zone_map(manifest: Mapping[str, Any]) -> dict[str, tuple[tuple[int, int, int, int], ...]]:
    result: dict[str, tuple[tuple[int, int, int, int], ...]] = {}
    for sheet in manifest.get("sheets") or []:
        result[str(sheet["sheet"])] = tuple(
            range_boundaries(str(zone["target"])) for zone in sheet.get("writable_zones") or []
        )
    return result


def _exact_writable_cells(bindings: Sequence[Mapping[str, Any]]) -> dict[str, set[str]]:
    result: dict[str, set[str]] = {}
    for binding in bindings:
        if not bool(binding.get("writable")) or str(binding.get("planning_state") or "active") != "active":
            continue
        sheet = str(binding.get("sheet") or "")
        target = str(binding.get("planner_target") or binding.get("target") or "")
        if not sheet or not target:
            continue
        min_col, min_row, max_col, max_row = range_boundaries(target)
        target_columns = [
            str(column.get("target_column") or "")
            for column in binding.get("target_columns") or []
            if isinstance(column, Mapping) and column.get("target_column")
        ]
        if target_columns:
            columns = [range_boundaries(f"{column}1")[0] for column in target_columns]
        else:
            columns = list(range(min_col, max_col + 1))
        cells = result.setdefault(sheet, set())
        for row in range(min_row, max_row + 1):
            for column in columns:
                if min_col <= column <= max_col:
                    cells.add(_coordinate(column, row))
    return result


def _canonical_color(color: Any) -> dict[str, Any] | None:
    if color is None:
        return None
    return {
        "type": str(getattr(color, "type", "") or ""),
        "rgb": str(getattr(color, "rgb", "") or ""),
        "indexed": getattr(color, "indexed", None),
        "auto": bool(getattr(color, "auto", False)),
        "theme": getattr(color, "theme", None),
        "tint": float(getattr(color, "tint", 0.0) or 0.0),
    }


def _canonical_side(side: Any) -> dict[str, Any]:
    return {
        "style": str(getattr(side, "style", "") or ""),
        "color": _canonical_color(getattr(side, "color", None)),
    }


def _canonical_fill(fill: Any) -> dict[str, Any]:
    stops = []
    for stop in getattr(fill, "stop", ()) or ():
        stops.append(
            {
                "position": float(getattr(stop, "position", 0.0) or 0.0),
                "color": _canonical_color(getattr(stop, "color", None)),
            }
        )
    return {
        "type": str(getattr(fill, "fill_type", "") or getattr(fill, "type", "") or ""),
        "fg_color": _canonical_color(getattr(fill, "fgColor", None)),
        "bg_color": _canonical_color(getattr(fill, "bgColor", None)),
        "degree": float(getattr(fill, "degree", 0.0) or 0.0),
        "left": float(getattr(fill, "left", 0.0) or 0.0),
        "right": float(getattr(fill, "right", 0.0) or 0.0),
        "top": float(getattr(fill, "top", 0.0) or 0.0),
        "bottom": float(getattr(fill, "bottom", 0.0) or 0.0),
        "stops": stops,
    }


def _canonical_cell_style(cell: Any) -> dict[str, Any]:
    font = cell.font
    border = cell.border
    alignment = cell.alignment
    protection = cell.protection
    return {
        "font": {
            "name": str(font.name or ""),
            "size": float(font.sz) if font.sz is not None else None,
            "bold": bool(font.b),
            "italic": bool(font.i),
            "underline": str(font.u or ""),
            "strike": bool(font.strike),
            "color": _canonical_color(font.color),
            "vert_align": str(font.vertAlign or ""),
            "outline": bool(font.outline),
            "shadow": bool(font.shadow),
            "condense": bool(font.condense),
            "extend": bool(font.extend),
            "scheme": str(font.scheme or ""),
            "family": font.family,
            "charset": font.charset,
        },
        "fill": _canonical_fill(cell.fill),
        "border": {
            "left": _canonical_side(border.left),
            "right": _canonical_side(border.right),
            "top": _canonical_side(border.top),
            "bottom": _canonical_side(border.bottom),
            "diagonal": _canonical_side(border.diagonal),
            "vertical": _canonical_side(border.vertical),
            "horizontal": _canonical_side(border.horizontal),
            "diagonal_up": bool(border.diagonalUp),
            "diagonal_down": bool(border.diagonalDown),
            "outline": bool(border.outline),
        },
        "alignment": {
            "horizontal": str(alignment.horizontal or ""),
            "vertical": str(alignment.vertical or ""),
            "text_rotation": int(alignment.textRotation or 0),
            "wrap_text": bool(alignment.wrapText),
            "shrink_to_fit": bool(alignment.shrinkToFit),
            "indent": float(alignment.indent or 0.0),
            "relative_indent": float(alignment.relativeIndent or 0.0),
            "justify_last_line": bool(alignment.justifyLastLine),
            "reading_order": float(alignment.readingOrder or 0.0),
        },
        "number_format": str(cell.number_format or "General"),
        "protection": {
            "locked": True if protection.locked is None else bool(protection.locked),
            "hidden": bool(protection.hidden),
        },
        "quote_prefix": bool(getattr(cell, "quotePrefix", False)),
        "pivot_button": bool(getattr(cell, "pivotButton", False)),
    }


_SIMPLE_QUOTED_SHEET_RE = re.compile(r"'([A-Za-z_][A-Za-z0-9_.]*)'!")


def _canonical_formula(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    is_formula = value.startswith("=")
    candidate = value if is_formula else f"={value}"
    try:
        tokens = Tokenizer(candidate).items
    except Exception:
        return _SIMPLE_QUOTED_SHEET_RE.sub(r"\1!", value)
    pieces: list[str] = []
    for token in tokens:
        token_value = str(token.value)
        if token.type == "OPERAND" and token.subtype == "NUMBER":
            token_value = _canonical_number_token(token_value)
        elif token.type == "OPERAND" and token.subtype == "RANGE":
            token_value = _SIMPLE_QUOTED_SHEET_RE.sub(r"\1!", token_value)
        elif token.type == "FUNC" and token.subtype == "OPEN" and token_value.endswith("("):
            token_value = f"{token_value[:-1].upper()}("
        pieces.append(token_value)
    result = "".join(pieces)
    return f"={result}" if is_formula else result


def _canonical_number_token(value: str) -> str:
    try:
        number = Decimal(value)
    except Exception:
        return value
    if number == number.to_integral():
        return str(number.quantize(Decimal("1")))
    return format(number.normalize(), "f").rstrip("0").rstrip(".")


def _quantize_dimension(value: Any, *, step: str) -> float | None:
    if value is None:
        return None
    quantum = Decimal(step)
    normalized = (Decimal(str(value)) / quantum).quantize(Decimal("1"), rounding=ROUND_HALF_UP) * quantum
    return float(normalized)


def _post_fill_structural_payload(wb: Any, *, allowed_cells: Mapping[str, set[str]]) -> dict[str, Any]:
    resolved_ticker_sheets = _resolved_ticker_sheet_names(wb)
    token_by_actual = {actual: token for token, actual in resolved_ticker_sheets.items()}
    sheet_name_map = {ws.title: token_by_actual.get(ws.title, ws.title) for ws in wb.worksheets}

    def normalized_text(value: Any) -> Any:
        if isinstance(value, str):
            for token, actual in sorted(resolved_ticker_sheets.items(), key=lambda item: len(item[1]), reverse=True):
                value = value.replace(f"'{actual}'", f"'{token}'").replace(actual, token)
        return value

    sheets = [{"sheet": sheet_name_map[ws.title], "state": ws.sheet_state} for ws in wb.worksheets]
    merges = sorted(
        f"{sheet_name_map[ws.title]}!{merged}"
        for ws in wb.worksheets
        for merged in ws.merged_cells.ranges
    )
    defined_names = sorted(
        (
            {
                "name": str(name),
                "attr_text": str(_canonical_formula(normalized_text(getattr(wb.defined_names[name], "attr_text", "") or ""))),
                "type": str(getattr(wb.defined_names[name], "type", "") or ""),
            }
            for name in wb.defined_names
        ),
        key=lambda row: (row["name"], row["attr_text"], row["type"]),
    )
    layout: list[dict[str, Any]] = []
    cell_structure: list[dict[str, Any]] = []
    data_validations: list[dict[str, Any]] = []
    conditional_formatting: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        normalized_sheet = sheet_name_map[ws.title]
        layout.append(
            {
                "sheet": normalized_sheet,
                "freeze_panes": str(ws.freeze_panes or ""),
                "sheet_format": {
                    "default_row_height": _quantize_dimension(ws.sheet_format.defaultRowHeight, step="0.05"),
                    "default_column_width": _quantize_dimension(ws.sheet_format.defaultColWidth, step="0.00390625"),
                },
                "rows": [
                    {
                        "index": int(index),
                        "height": _quantize_dimension(dimension.height, step="0.05"),
                        "hidden": bool(dimension.hidden),
                        "outline": int(dimension.outlineLevel or 0),
                    }
                    for index, dimension in sorted(ws.row_dimensions.items())
                    if dimension.height is not None or dimension.hidden or dimension.outlineLevel
                ],
                "columns": [
                    {
                        "key": str(key),
                        "width": _quantize_dimension(dimension.width, step="0.00390625"),
                        "hidden": bool(dimension.hidden),
                        "outline": int(dimension.outlineLevel or 0),
                    }
                    for key, dimension in sorted(ws.column_dimensions.items())
                    if dimension.width is not None or dimension.hidden or dimension.outlineLevel
                ],
            }
        )
        allowed = allowed_cells.get(normalized_sheet, set())
        # Materialize every declared writable cell in both source and target so
        # assigning a value to a previously absent blank cell does not look like
        # a structural cell-collection change.
        for coordinate in allowed:
            ws[coordinate]
        for cell in sorted(ws._cells.values(), key=lambda item: (item.row, item.column)):
            is_writable = cell.coordinate in allowed
            value = normalized_text(cell.value)
            cell_structure.append(
                {
                    "sheet": normalized_sheet,
                    "cell": cell.coordinate,
                    "style": _canonical_cell_style(cell),
                    "comment": (
                        {"author": str(cell.comment.author or ""), "text": str(cell.comment.text or "")}
                        if cell.comment is not None
                        else None
                    ),
                    "writable": is_writable,
                    "writable_formula": bool(is_writable and isinstance(value, str) and value.startswith("=")),
                    "protected_value": (
                        None
                        if is_writable
                        else _canonical_formula(value)
                        if isinstance(value, str) and value.startswith("=")
                        else value
                    ),
                }
            )
        for validation in ws.data_validations.dataValidation:
            data_validations.append(
                {
                    "sheet": normalized_sheet,
                    "sqref": str(validation.sqref),
                    "type": str(validation.type or ""),
                    "operator": str(validation.operator or ""),
                    "formula1": str(_canonical_formula(normalized_text(validation.formula1)) or ""),
                    "formula2": str(_canonical_formula(normalized_text(validation.formula2)) or ""),
                }
            )
        for conditional_range in ws.conditional_formatting:
            rules = ws.conditional_formatting[conditional_range]
            conditional_formatting.append(
                {
                    "sheet": normalized_sheet,
                    "sqref": str(conditional_range),
                    "rules": [
                        {
                            "type": str(rule.type or ""),
                            "operator": str(rule.operator or ""),
                            "formula": [str(_canonical_formula(normalized_text(value)) or "") for value in (rule.formula or [])],
                            "priority": int(rule.priority or 0),
                        }
                        for rule in rules
                    ],
                }
            )
        for table_name in sorted(ws.tables):
            table = ws.tables[table_name]
            tables.append(
                {
                    "sheet": normalized_sheet,
                    "name": str(table.name),
                    "display_name": str(table.displayName),
                    "ref": str(table.ref),
                }
            )
    return {
        "sheets": sheets,
        "sheet_views": _sheet_view_contract_payload(wb, sheet_name_map=sheet_name_map),
        "merges": merges,
        "defined_names": defined_names,
        "layout": layout,
        "cell_structure": cell_structure,
        "data_validations": sorted(data_validations, key=lambda row: (row["sheet"], row["sqref"])),
        "conditional_formatting": sorted(conditional_formatting, key=lambda row: (row["sheet"], row["sqref"])),
        "tables": sorted(tables, key=lambda row: (row["sheet"], row["name"])),
    }


def _writable_value_map(wb: Any, *, allowed_cells: Mapping[str, set[str]]) -> dict[str, Any]:
    resolved = _resolved_ticker_sheet_names(wb)
    result: dict[str, Any] = {}
    for template_sheet, cells in allowed_cells.items():
        actual_sheet = resolved.get(template_sheet, template_sheet)
        if actual_sheet not in wb.sheetnames:
            continue
        ws = wb[actual_sheet]
        for coordinate in cells:
            result[f"{template_sheet}!{coordinate}"] = ws[coordinate].value
    return result


def _planned_value_map(plan: Mapping[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for write in plan.get("planned_writes") or []:
        if not isinstance(write, Mapping):
            continue
        sheet = str(write.get("target_sheet") or "")
        if sheet.endswith("_Investment_Case_Data"):
            sheet = "{ticker}_Investment_Case_Data"
        elif sheet.endswith("_Investment_Case"):
            sheet = "{ticker}_Investment_Case"
        key = f"{sheet}!{write.get('target_cell') or ''}"
        result[key] = write.get("value")
    return result


def _validate_plan_shell_identity(
    plan: Mapping[str, Any],
    source_token: VerifiedShellIdentity,
    issues: list[dict[str, str]],
) -> None:
    plan_identity = plan.get("shell_identity") if isinstance(plan.get("shell_identity"), Mapping) else {}
    plan_actual = plan_identity.get("actual") if isinstance(plan_identity.get("actual"), Mapping) else {}
    source_actual = source_token.actual
    fields = ("semantic_contract_version", *IDENTITY_FIELDS)
    if any(str(plan_actual.get(field) or "") != str(source_actual.get(field) or "") for field in fields):
        issues.append(
            {
                "rule_id": "post_fill_plan_shell_identity_mismatch",
                "message": "The approved binding plan was not issued for the verified source shell identity.",
            }
        )


def _resolved_ticker_sheet_name(wb: Any) -> str:
    return _resolved_ticker_sheet_names(wb).get("{ticker}_Investment_Case", "")


def _resolved_ticker_sheet_names(wb: Any) -> dict[str, str]:
    resolved: dict[str, str] = {}
    for token in ("{ticker}_Investment_Case", "{ticker}_Investment_Case_Data"):
        if token in wb.sheetnames:
            resolved[token] = token
            continue
        suffix = token.replace("{ticker}", "")
        candidates = [name for name in wb.sheetnames if name.endswith(suffix)]
        if len(candidates) == 1:
            resolved[token] = candidates[0]
    return resolved


def _coordinate(column: int, row: int) -> str:
    letters = ""
    current = column
    while current:
        current, remainder = divmod(current - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{row}"


def _cell_in_zones(column: int, row: int, zones: Sequence[tuple[int, int, int, int]]) -> bool:
    return any(min_col <= column <= max_col and min_row <= row <= max_row for min_col, min_row, max_col, max_row in zones)


def _writable_target_contract(manifest: Mapping[str, Any], bindings: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    contracts = [
        {
            "contract_id": str(item.get("contract_id") or ""),
            "sheet": str(item.get("sheet") or ""),
            "target": str(item.get("target") or ""),
            "target_role": str(item.get("target_role") or ""),
            "allowed_binding_ids": sorted(str(value) for value in item.get("allowed_binding_ids") or []),
            "allowed_target_types": sorted(str(value) for value in item.get("allowed_target_types") or []),
        }
        for item in manifest.get("planner_cell_contracts") or []
        if bool(item.get("writable"))
    ]
    binding_targets = [
        {
            "binding_id": str(binding.get("binding_id") or ""),
            "sheet": str(binding.get("sheet") or ""),
            "planner_target": str(binding.get("planner_target") or binding.get("target") or ""),
            "planning_mode": str(binding.get("planning_mode") or ""),
            "target_columns": [
                {
                    "source_field": str(column.get("source_field") or ""),
                    "target_column": str(column.get("target_column") or ""),
                    "target_type": str(column.get("target_type") or ""),
                    "target_role": str(column.get("target_role") or ""),
                }
                for column in binding.get("target_columns") or []
            ],
        }
        for binding in bindings
        if bool(binding.get("writable")) and str(binding.get("planning_state") or "active") == "active"
    ]
    return sorted([*contracts, *binding_targets], key=lambda row: json.dumps(row, sort_keys=True))


def _executable_binding_contract(bindings: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    """Canonicalize the full binding contract without depending on JSON order."""

    contracts = [_contract_value(binding) for binding in bindings if isinstance(binding, Mapping)]
    return sorted(
        contracts,
        key=lambda row: (
            str(row.get("binding_id") or ""),
            json.dumps(row, sort_keys=True, ensure_ascii=False, separators=(",", ":")),
        ),
    )


def _binding_document_contract(
    binding_payload: Mapping[str, Any] | Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    if isinstance(binding_payload, Mapping):
        return _contract_value(binding_payload)
    return {
        "version": "legacy-sequence-contract",
        "binding_planner_contract_version": BINDING_PLANNER_CONTRACT_VERSION,
        "bindings": _executable_binding_contract(binding_payload),
    }


def _contract_value(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _contract_value(child) for key, child in sorted(value.items(), key=lambda row: str(row[0]))}
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        return [_contract_value(child) for child in value]
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _planned_cell_values_equal(actual: Any, expected: Any) -> bool:
    """Compare a saved cell with its plan without masking business drift.

    Excel/Open XML writers may round an IEEE-754 value at the last represented
    digit.  Only finite, non-boolean numeric scalars receive this narrow relative
    tolerance; strings, formulas, dates, booleans, and structures remain exact.
    """

    if actual == expected:
        return True
    if (
        isinstance(actual, (int, float))
        and not isinstance(actual, bool)
        and isinstance(expected, (int, float))
        and not isinstance(expected, bool)
    ):
        actual_number = float(actual)
        expected_number = float(expected)
        return math.isfinite(actual_number) and math.isfinite(expected_number) and math.isclose(
            actual_number,
            expected_number,
            rel_tol=2e-15,
            abs_tol=0.0,
        )
    return _contract_value(actual) == _contract_value(expected)


def _non_writable_contract(manifest: Mapping[str, Any]) -> list[dict[str, str]]:
    return sorted(
        (
            {
                "sheet": str(sheet["sheet"]),
                "zone_id": str(zone["zone_id"]),
                "target": str(zone["target"]),
            }
            for sheet in manifest.get("sheets") or []
            for zone in sheet.get("non_writable_zones") or []
        ),
        key=lambda row: (row["sheet"], row["zone_id"], row["target"]),
    )


def _signature(value: Any) -> str:
    canonical = json.dumps(value, sort_keys=True, ensure_ascii=False, separators=(",", ":"), default=str)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _identity_rule_id(field: str) -> str:
    return {
        "workbook_sha256": "shell_workbook_sha256_mismatch",
        "manifest_contract_signature": "shell_manifest_contract_drift",
        "sheet_order_visibility_signature": "shell_sheet_order_visibility_drift",
        "sheet_view_signature": "shell_sheet_view_drift",
        "merge_signature": "shell_merge_drift",
        "defined_name_signature": "shell_defined_name_drift",
        "writable_target_signature": "shell_writable_target_drift",
        "binding_contract_signature": "shell_binding_contract_drift",
        "formula_static_zone_signature": "shell_formula_static_zone_drift",
    }[field]


def _contract_schema_issues(
    document: Mapping[str, Any],
    schema_path: Path,
    *,
    rule_prefix: str,
) -> list[dict[str, str]]:
    try:
        schema = load_json_strict(schema_path)
        failures = validate_json_schema(document, schema)
    except Exception as exc:
        return [{"rule_id": f"{rule_prefix}_unavailable", "message": str(exc)}]
    return [
        {
            "rule_id": f"{rule_prefix}_{keyword}",
            "message": f"{path}: {message}",
        }
        for path, keyword, message in failures
    ]
