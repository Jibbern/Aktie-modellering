"""Workbook and cache comparison helpers for gap-audit reporting."""
from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook

from .company_profiles import get_company_profile
from .excel_writer import build_saved_workbook_provenance
from .excel_writer_segments import (
    latest_segment_financials_workbook,
    parse_quarterly_segment_data_from_workbook,
)


PIPELINE_BUNDLE_FIELD_NAMES: Tuple[str, ...] = (
    "hist",
    "audit",
    "debt_tranches",
    "debt_recon",
    "adj_metrics",
    "adj_breakdown",
    "non_gaap_files",
    "adj_metrics_relaxed",
    "adj_breakdown_relaxed",
    "non_gaap_files_relaxed",
    "needs_review",
    "info_log",
    "tag_coverage",
    "period_checks",
    "qa_checks",
    "bridge_q",
    "manifest_df",
    "ocr_log",
    "qfd_preview",
    "qfd_unused",
    "debt_profile",
    "debt_tranches_latest",
    "debt_maturity",
    "debt_credit_notes",
    "revolver_df",
    "revolver_history",
    "debt_buckets",
    "slides_segments",
    "slides_debt",
    "slides_guidance",
    "quarter_notes",
    "promises",
    "promise_progress",
    "non_gaap_cred",
    "company_overview",
)

STATUS_NOT_FOUND = "not_found"
STATUS_PARSE_FAILED = "source_exists_parse_failed"
STATUS_WRITE_FAILED = "source_exists_write_failed"


@dataclass(frozen=True)
class GapMatrixRow:
    ticker: str
    sheet: str
    row_label: str
    period: str
    expected_source_family: str
    status: str
    matched_source_path: str
    reason: str


def _bundle_path(repo_root: Path, ticker: str) -> Path:
    ticker_u = str(ticker or "").strip().upper()
    return repo_root / "sec_cache" / ticker_u / "pipeline_bundle_cache" / f"{ticker_u}.pkl"


def load_pipeline_bundle_map(repo_root: Path, ticker: str) -> Dict[str, Any]:
    bundle_path = _bundle_path(repo_root, ticker)
    obj = pd.read_pickle(bundle_path)
    if isinstance(obj, tuple):
        return {
            name: obj[idx] if idx < len(obj) else None
            for idx, name in enumerate(PIPELINE_BUNDLE_FIELD_NAMES)
        }
    if hasattr(obj, "__dict__"):
        return {name: getattr(obj, name, None) for name in PIPELINE_BUNDLE_FIELD_NAMES}
    raise TypeError(f"Unsupported pipeline bundle type for {ticker}: {type(obj)!r}")


def classify_gap_status(*, source_exists: bool, parse_ok: bool, write_ok: bool) -> str:
    if not source_exists:
        return STATUS_NOT_FOUND
    if not parse_ok:
        return STATUS_PARSE_FAILED
    if not write_ok:
        return STATUS_WRITE_FAILED
    raise ValueError("Gap status requested for a non-gap case.")


def infer_expected_source_family(*, sheet: str, metric: str = "", issue_family: str = "", row_label: str = "") -> str:
    key = " ".join([str(sheet or ""), str(metric or ""), str(issue_family or ""), str(row_label or "")]).lower()
    if any(tok in key for tok in ("convertible", "debt", "revolver", "leverage", "liquidity")):
        return "debt_tranches / debt_recon / tranche-sources"
    if "segment" in key:
        return "historical_segment / segment_financials"
    if any(tok in key for tok in ("quarter_notes_ui", "promise_progress_ui", "promise", "note")):
        return "earnings_release / earnings_presentation / earnings_transcripts"
    return "earnings_release / earnings_presentation / earnings_transcripts"


def write_workbook_gap_matrix_report(
    repo_root: Path,
    *,
    tickers: Sequence[str] = ("PBI", "GPRE"),
    output_path: Optional[Path] = None,
) -> Path:
    rows: List[GapMatrixRow] = []
    for ticker in tickers:
        rows.extend(build_workbook_gap_rows(repo_root, ticker))
    out_path = output_path or (repo_root / "sec_cache" / "_reports" / "workbook_gap_matrix_latest.md")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(render_workbook_gap_matrix_markdown(rows), encoding="utf-8")
    return out_path


def render_workbook_gap_matrix_markdown(rows: Sequence[GapMatrixRow]) -> str:
    lines: List[str] = [
        "# Workbook Gap Matrix",
        "",
        "Status values:",
        f"- `{STATUS_NOT_FOUND}`",
        f"- `{STATUS_PARSE_FAILED}`",
        f"- `{STATUS_WRITE_FAILED}`",
        "",
    ]
    by_ticker: Dict[str, List[GapMatrixRow]] = {}
    for row in rows:
        by_ticker.setdefault(row.ticker, []).append(row)
    for ticker in sorted(by_ticker):
        lines.append(f"## {ticker}")
        lines.append("")
        lines.append("| Sheet | Row / metric | Period | Expected source family | Status | Matched source path | Reason |")
        lines.append("| --- | --- | --- | --- | --- | --- | --- |")
        ticker_rows = sorted(
            by_ticker[ticker],
            key=lambda r: (r.sheet, r.row_label, r.period, r.status, r.reason),
        )
        if not ticker_rows:
            lines.append("| - | - | - | - | - | - | - |")
        for row in ticker_rows:
            lines.append(
                "| "
                + " | ".join(
                    [
                        _md_cell(row.sheet),
                        _md_cell(row.row_label),
                        _md_cell(row.period),
                        _md_cell(row.expected_source_family),
                        _md_cell(row.status),
                        _md_cell(row.matched_source_path),
                        _md_cell(row.reason),
                    ]
                )
                + " |"
            )
        lines.append("")
    return "\n".join(lines)


def build_workbook_gap_rows(repo_root: Path, ticker: str) -> List[GapMatrixRow]:
    ticker_u = str(ticker or "").strip().upper()
    workbook_path = repo_root / "Excel stock models" / f"{ticker_u}_model.xlsx"
    bundle = load_pipeline_bundle_map(repo_root, ticker_u)
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        rows: List[GapMatrixRow] = []
        rows.extend(_collect_required_sheet_gaps(ticker_u, wb))
        rows.extend(_collect_needs_review_gaps(repo_root, ticker_u, wb, bundle))
        rows.extend(_collect_segment_gaps(repo_root, ticker_u, wb))
        rows.extend(_collect_convertible_gaps(ticker_u, wb, bundle))
        rows.extend(_collect_quarter_note_and_promise_gaps(ticker_u, workbook_path, bundle))
        return rows
    finally:
        wb.close()


def _collect_required_sheet_gaps(ticker: str, wb: Any) -> List[GapMatrixRow]:
    required = {
        "PBI": ["Operating_Drivers", "Quarter_Notes_UI", "Promise_Progress_UI", "Needs_Review", "QA_Log", "Valuation", "BS_Segments"],
        "GPRE": ["Operating_Drivers", "Quarter_Notes_UI", "Promise_Progress_UI", "Needs_Review", "QA_Log", "Valuation", "BS_Segments", "Economics_Overlay"],
    }.get(ticker, [])
    rows: List[GapMatrixRow] = []
    for sheet_name in required:
        if sheet_name not in wb.sheetnames:
            rows.append(
                GapMatrixRow(
                    ticker=ticker,
                    sheet=sheet_name,
                    row_label="sheet_missing",
                    period="all_visible_history",
                    expected_source_family=infer_expected_source_family(sheet=sheet_name),
                    status=STATUS_NOT_FOUND,
                    matched_source_path="",
                    reason="Required visible sheet is missing from saved workbook.",
                )
            )
    return rows


def _collect_needs_review_gaps(repo_root: Path, ticker: str, wb: Any, bundle: Dict[str, Any]) -> List[GapMatrixRow]:
    if "Needs_Review" not in wb.sheetnames:
        return []
    ws = wb["Needs_Review"]
    header = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    col_map = {name: idx + 1 for idx, name in enumerate(header) if name}
    rows: List[GapMatrixRow] = []
    latest_q = _latest_quarter_from_hist(bundle.get("hist"))
    latest_q_txt = latest_q.isoformat() if isinstance(latest_q, date) else ""
    latest_issue_count = 0
    for rr in range(2, ws.max_row + 1):
        issue_family = str(_ws_value(ws, rr, col_map.get("issue_family")) or "").strip()
        latest_message = str(_ws_value(ws, rr, col_map.get("latest_message")) or "").strip()
        period_txt = _dateish_to_iso(_ws_value(ws, rr, col_map.get("last_seen_q")) or _ws_value(ws, rr, col_map.get("quarter")))
        source_txt = str(_ws_value(ws, rr, col_map.get("source")) or "").strip()
        row_label = issue_family or str(_ws_value(ws, rr, col_map.get("raw_metric")) or "").strip() or "Needs_Review"
        if not row_label and not latest_message:
            continue
        if latest_q_txt and period_txt == latest_q_txt:
            latest_issue_count += 1
        source_exists = bool(source_txt)
        parse_ok = not any(
            tok in issue_family
            for tok in (
                "quarter_text_no_explicit_support",
                "quarter_text_low_confidence_support",
            )
        )
        write_ok = not any(
            tok in issue_family
            for tok in (
                "segment",
                "debt_recon_coverage_check",
            )
        )
        if issue_family == "debt_recon_coverage_check":
            dtl = bundle.get("debt_tranches_latest")
            source_exists = not _latest_quarter_debt_rows(dtl, latest_q).empty
            parse_ok = source_exists
            write_ok = False
        elif issue_family == "segment":
            seg_path, seg_parsed = _load_segment_workbook_parse(repo_root=repo_root, ticker=ticker)
            source_exists = seg_path is not None
            parse_ok = bool(seg_parsed.get("metrics")) if isinstance(seg_parsed, dict) else False
            write_ok = False
        elif issue_family == "quarter_text_numeric_conflict":
            source_exists = bool(source_txt)
            parse_ok = True
            write_ok = False
        if source_exists and parse_ok and write_ok:
            write_ok = False
        status = classify_gap_status(
            source_exists=source_exists,
            parse_ok=parse_ok,
            write_ok=write_ok,
        )
        rows.append(
            GapMatrixRow(
                ticker=ticker,
                sheet="Needs_Review",
                row_label=row_label,
                period=period_txt or "unknown",
                expected_source_family=infer_expected_source_family(sheet="Needs_Review", issue_family=issue_family, row_label=row_label),
                status=status,
                matched_source_path=source_txt,
                reason=latest_message,
            )
        )
    if latest_issue_count == 0 and latest_q_txt:
        rows.append(
            GapMatrixRow(
                ticker=ticker,
                sheet="Needs_Review",
                row_label="latest_quarter_review_rows",
                period=latest_q_txt,
                expected_source_family="QA / curated review queue",
                status=STATUS_WRITE_FAILED,
                matched_source_path="Needs_Review",
                reason="No latest-quarter items surfaced in Needs_Review despite visible QA issues being expected.",
            )
        )
    return rows


def _collect_segment_gaps(repo_root: Path, ticker: str, wb: Any) -> List[GapMatrixRow]:
    if "Operating_Drivers" not in wb.sheetnames:
        return []
    seg_path, parsed = _load_segment_workbook_parse(repo_root=repo_root, ticker=ticker)
    expected_family = infer_expected_source_family(sheet="Operating_Drivers", row_label="segment_support")
    if seg_path is None:
        return [
            GapMatrixRow(
                ticker=ticker,
                sheet="Operating_Drivers",
                row_label="segment_support_bundle",
                period="latest_quarter",
                expected_source_family=expected_family,
                status=STATUS_NOT_FOUND,
                matched_source_path="",
                reason="No quarterly segment workbook was found under ticker-root segment directories.",
            )
        ]
    if not parsed.get("metrics") or not parsed.get("quarters"):
        return [
            GapMatrixRow(
                ticker=ticker,
                sheet="Operating_Drivers",
                row_label="segment_support_bundle",
                period="latest_quarter",
                expected_source_family=expected_family,
                status=STATUS_PARSE_FAILED,
                matched_source_path=str(seg_path),
                reason="Segment workbook exists but quarterly segment parsing returned no usable metrics.",
            )
        ]
    latest_q = max(pd.Timestamp(qd).date() for qd in parsed.get("quarters") or [])
    ws = wb["Operating_Drivers"]
    q_col_map = _operating_drivers_quarter_cols(ws)
    rows_out: List[GapMatrixRow] = []
    latest_key = f"{latest_q.year}-Q{pd.Timestamp(latest_q).quarter}"
    target_col = q_col_map.get(latest_key)
    if target_col is None:
        rows_out.append(
            GapMatrixRow(
                ticker=ticker,
                sheet="Operating_Drivers",
                row_label="segment_support_quarter_headers",
                period=latest_key,
                expected_source_family=expected_family,
                status=STATUS_WRITE_FAILED,
                matched_source_path=str(seg_path),
                reason="Operating_Drivers is missing the latest-quarter column used by the segment support block.",
            )
        )
        return rows_out
    profile = get_company_profile(ticker)
    metric_sections = {
        "Revenue": ("Revenue ($m)", 1_000_000.0),
        "Adjusted EBIT": ("Adj EBIT / operating profit ($m)", 1_000_000.0),
    }
    for metric_name, (section_label, scale) in metric_sections.items():
        section_rows = _operating_drivers_section_rows(ws, section_label)
        metric_store = dict(parsed.get("metrics", {}).get(metric_name) or {})
        for seg_name in getattr(profile, "quarterly_segment_labels", tuple()) or tuple(metric_store.keys()):
            seg_series = dict(metric_store.get(str(seg_name)) or {})
            if pd.Timestamp(latest_q) not in seg_series:
                continue
            expected_val = float(seg_series[pd.Timestamp(latest_q)]) / float(scale)
            row_idx = section_rows.get(str(seg_name))
            if row_idx is None:
                rows_out.append(
                    GapMatrixRow(
                        ticker=ticker,
                        sheet="Operating_Drivers",
                        row_label=f"{section_label} | {seg_name}",
                        period=latest_key,
                        expected_source_family=expected_family,
                        status=STATUS_WRITE_FAILED,
                        matched_source_path=str(seg_path),
                        reason="Parsed segment source exists, but the segment row is missing from Operating_Drivers.",
                    )
                )
                continue
            actual_val = pd.to_numeric(ws.cell(row_idx, target_col).value, errors="coerce")
            if pd.isna(actual_val):
                rows_out.append(
                    GapMatrixRow(
                        ticker=ticker,
                        sheet="Operating_Drivers",
                        row_label=f"{section_label} | {seg_name}",
                        period=latest_key,
                        expected_source_family=expected_family,
                        status=STATUS_WRITE_FAILED,
                        matched_source_path=str(seg_path),
                        reason="Parsed segment source exists, but the latest-quarter cell is blank in Operating_Drivers.",
                    )
                )
                continue
            diff_abs = abs(float(actual_val) - expected_val)
            if diff_abs > max(0.05, abs(expected_val) * 0.01):
                rows_out.append(
                    GapMatrixRow(
                        ticker=ticker,
                        sheet="Operating_Drivers",
                        row_label=f"{section_label} | {seg_name}",
                        period=latest_key,
                        expected_source_family=expected_family,
                        status=STATUS_WRITE_FAILED,
                        matched_source_path=str(seg_path),
                        reason=f"Parsed segment value {expected_val:,.3f} vs saved workbook {float(actual_val):,.3f}.",
                    )
                )
    return rows_out


def _collect_convertible_gaps(ticker: str, wb: Any, bundle: Dict[str, Any]) -> List[GapMatrixRow]:
    if "Valuation" not in wb.sheetnames:
        return []
    ws = wb["Valuation"]
    hdr_row = _find_value_in_column(ws, "Convertible notes", 12)
    if hdr_row is None:
        return []
    latest_q = _latest_quarter_from_hist(bundle.get("hist"))
    latest_key = f"{latest_q.year}-Q{pd.Timestamp(latest_q).quarter}" if isinstance(latest_q, date) else "latest_quarter"
    dtl = bundle.get("debt_tranches_latest")
    raw_dt = bundle.get("debt_tranches")
    latest_dtl = _latest_quarter_debt_rows(dtl, latest_q)
    raw_convert = _latest_quarter_convertible_rows(raw_dt, latest_q)
    visible_marker = str(ws.cell(hdr_row + 2, 12).value or "").strip()
    rows: List[GapMatrixRow] = []
    if not raw_convert.empty and latest_dtl.empty and "No convertible debt identified" in visible_marker:
        source_doc = " | ".join(sorted({str(v) for v in raw_convert.get("doc", pd.Series(dtype=object)).dropna().tolist() if str(v).strip()}))
        rows.append(
            GapMatrixRow(
                ticker=ticker,
                sheet="Valuation",
                row_label="Convertible notes",
                period=latest_key,
                expected_source_family=infer_expected_source_family(sheet="Valuation", row_label="Convertible notes"),
                status=STATUS_PARSE_FAILED,
                matched_source_path=source_doc,
                reason="Raw latest-quarter tranche rows include convertible debt, but the latest debt view is empty so the visible convertible block is blank.",
            )
        )
        return rows
    if latest_dtl.empty:
        return rows
    visible_rows = _visible_convertible_rows(ws, hdr_row + 2)
    if not visible_rows:
        source_doc = " | ".join(sorted({str(v) for v in latest_dtl.get("doc", pd.Series(dtype=object)).dropna().tolist() if str(v).strip()}))
        rows.append(
            GapMatrixRow(
                ticker=ticker,
                sheet="Valuation",
                row_label="Convertible notes",
                period=latest_key,
                expected_source_family=infer_expected_source_family(sheet="Valuation", row_label="Convertible notes"),
                status=STATUS_WRITE_FAILED,
                matched_source_path=source_doc,
                reason="Latest-quarter convertible tranches exist, but no security rows were written in the visible convertible block.",
            )
        )
        return rows
    for rr in visible_rows:
        label_txt = str(ws.cell(rr, 12).value or "").strip()
        if not label_txt:
            continue
        matched = _match_convertible_latest_row(latest_dtl, label_txt)
        if matched is None:
            continue
        conv_price_val = _ws_value(ws, rr, 17)
        shares_val = _ws_value(ws, rr, 19)
        if _cell_has_visible_value(conv_price_val) and _cell_has_visible_value(shares_val):
            continue
        parsed_price = pd.to_numeric(matched.get("conversion_price"), errors="coerce")
        parsed_shares = pd.to_numeric(matched.get("shares_on_full_conversion"), errors="coerce")
        source_doc = str(
            matched.get("conversion_terms_source")
            or matched.get("doc")
            or matched.get("source")
            or ""
        ).strip()
        parse_ok = bool(pd.notna(parsed_price) or pd.notna(parsed_shares))
        missing_parts: List[str] = []
        if not _cell_has_visible_value(conv_price_val):
            missing_parts.append("conversion price")
        if not _cell_has_visible_value(shares_val):
            missing_parts.append("shares on full conversion")
        rows.append(
            GapMatrixRow(
                ticker=ticker,
                sheet="Valuation",
                row_label=label_txt,
                period=latest_key,
                expected_source_family=infer_expected_source_family(sheet="Valuation", row_label="Convertible notes"),
                status=STATUS_WRITE_FAILED if parse_ok else STATUS_PARSE_FAILED,
                matched_source_path=source_doc,
                reason=(
                    f"Visible convertible row is missing {', '.join(missing_parts)} "
                    f"for latest-quarter tranche '{label_txt}'."
                    if parse_ok
                    else f"Latest-quarter tranche '{label_txt}' was written, but parsed conversion terms are still missing."
                ),
            )
        )
    return rows


def _collect_quarter_note_and_promise_gaps(ticker: str, workbook_path: Path, bundle: Dict[str, Any]) -> List[GapMatrixRow]:
    provenance = build_saved_workbook_provenance(workbook_path)
    quarter_note_df = bundle.get("quarter_notes")
    promise_progress_df = bundle.get("promise_progress")
    rows: List[GapMatrixRow] = []
    qn_snapshot = dict(provenance.get("quarter_notes_ui_snapshot") or {})
    pp_quarters = set(_quarter_values_from_df(promise_progress_df))
    latest_hist_q = _latest_quarter_from_hist(bundle.get("hist"))
    latest_q_txt = latest_hist_q.isoformat() if isinstance(latest_hist_q, date) else ""
    if latest_q_txt:
        runtime_qn = _quarter_values_from_df(quarter_note_df)
        if latest_q_txt in runtime_qn and latest_q_txt not in qn_snapshot:
            rows.append(
                GapMatrixRow(
                    ticker=ticker,
                    sheet="Quarter_Notes_UI",
                    row_label="quarter_block",
                    period=latest_q_txt,
                    expected_source_family=infer_expected_source_family(sheet="Quarter_Notes_UI"),
                    status=STATUS_WRITE_FAILED,
                    matched_source_path="Quarter_Notes_UI",
                    reason="Quarter notes runtime contains latest-quarter evidence, but the visible quarter block is missing.",
                )
            )
        if latest_q_txt in pp_quarters and "Promise_Progress_UI" not in _visible_sheet_names_from_path(workbook_path):
            rows.append(
                GapMatrixRow(
                    ticker=ticker,
                    sheet="Promise_Progress_UI",
                    row_label="sheet_missing",
                    period=latest_q_txt,
                    expected_source_family=infer_expected_source_family(sheet="Promise_Progress_UI"),
                    status=STATUS_WRITE_FAILED,
                    matched_source_path="Promise_Progress_UI",
                    reason="Promise progress runtime has latest-quarter rows, but the visible sheet is missing.",
                )
            )
    return rows


def _load_segment_workbook_parse(repo_root: Path, ticker: str, bundle: Optional[Dict[str, Any]] = None) -> Tuple[Optional[Path], Dict[str, Any]]:
    ticker_root = repo_root / str(ticker or "").strip().upper()
    profile = get_company_profile(ticker)
    for dir_name in ("historical_segment", "segment_financials", "financial_statement"):
        workbook_path = latest_segment_financials_workbook(ticker_root / dir_name)
        if workbook_path is None:
            continue
        parsed = parse_quarterly_segment_data_from_workbook(
            workbook_path,
            annual_segment_alias_patterns=profile.annual_segment_alias_patterns,
            company_segment_alias_patterns=profile.segment_alias_patterns,
        )
        return workbook_path, parsed or {}
    return None, {}


def _latest_quarter_from_hist(hist: Any) -> Optional[date]:
    if hist is None or getattr(hist, "empty", True) or "quarter" not in hist.columns:
        return None
    q = pd.to_datetime(hist["quarter"], errors="coerce").dropna()
    if q.empty:
        return None
    return pd.Timestamp(q.max()).date()


def _latest_quarter_debt_rows(df: Any, latest_q: Optional[date]) -> pd.DataFrame:
    if df is None or getattr(df, "empty", True) or latest_q is None or "quarter" not in df.columns:
        return pd.DataFrame()
    out = df.copy()
    out["quarter"] = pd.to_datetime(out["quarter"], errors="coerce")
    out = out[out["quarter"].notna()]
    return out[out["quarter"].dt.to_period("Q") == pd.Timestamp(latest_q).to_period("Q")].copy()


def _latest_quarter_convertible_rows(df: Any, latest_q: Optional[date]) -> pd.DataFrame:
    if df is None or getattr(df, "empty", True):
        return pd.DataFrame()
    out = _latest_quarter_debt_rows(df, latest_q)
    if out.empty:
        return out
    name_col = "tranche_name" if "tranche_name" in out.columns else "name" if "name" in out.columns else None
    if name_col is None:
        return pd.DataFrame()
    return out[out[name_col].astype(str).str.contains("convert", case=False, na=False)].copy()


def _operating_drivers_quarter_cols(ws: Any) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for rr in range(1, min(ws.max_row, 60) + 1):
        if str(ws.cell(rr, 1).value or "").strip() != "Metric / segment":
            continue
        for cc in range(2, ws.max_column + 1):
            txt = str(ws.cell(rr, cc).value or "").strip()
            if txt:
                out[txt] = cc
        break
    return out


def _operating_drivers_section_rows(ws: Any, section_label: str) -> Dict[str, int]:
    rows: Dict[str, int] = {}
    section_row = _find_value_in_column(ws, section_label, 1)
    if section_row is None:
        return rows
    for rr in range(section_row + 1, ws.max_row + 1):
        label = str(ws.cell(rr, 1).value or "").strip()
        if not label:
            break
        if label in {"Revenue ($m)", "Adj EBIT / operating profit ($m)", "Margin"}:
            break
        rows[label] = rr
    return rows


def _find_value_in_column(ws: Any, needle: str, column: int) -> Optional[int]:
    needle_txt = str(needle or "").strip()
    for rr in range(1, ws.max_row + 1):
        if str(ws.cell(rr, column).value or "").strip() == needle_txt:
            return rr
    return None


def _visible_convertible_rows(ws: Any, start_row: int) -> List[int]:
    rows: List[int] = []
    for rr in range(start_row, min(ws.max_row, start_row + 10) + 1):
        label = str(ws.cell(rr, 12).value or "").strip()
        if not label:
            continue
        if label in {"Total"} or label.startswith("When valuing"):
            continue
        if label.startswith("No convertible debt identified"):
            continue
        rows.append(rr)
    return rows


def _cell_has_visible_value(value: Any) -> bool:
    if value is None:
        return False
    try:
        if pd.isna(value):
            return False
    except Exception:
        pass
    return str(value).strip() != ""


def _match_convertible_latest_row(df: pd.DataFrame, label_txt: str) -> Optional[pd.Series]:
    if df is None or getattr(df, "empty", True):
        return None
    label_low = str(label_txt or "").lower()
    label_norm = " ".join(label_low.replace("%", " % ").split())
    year_match = pd.to_numeric(re.search(r"(20\d{2})", label_low).group(1), errors="coerce") if re.search(r"(20\d{2})", label_low) else pd.NA
    coupon_match = re.search(r"([0-9]+(?:\.\d+)?)\s*%", label_low)
    coupon_num = pd.to_numeric(coupon_match.group(1), errors="coerce") if coupon_match else pd.NA
    best_row: Optional[pd.Series] = None
    best_score = -1
    for _, row in df.iterrows():
        score = 0
        row_name = str(row.get("tranche_name") or "").lower()
        row_norm = " ".join(row_name.replace("%", " % ").split())
        row_year = pd.to_numeric(row.get("maturity_year"), errors="coerce")
        row_coupon = pd.to_numeric(row.get("coupon_pct"), errors="coerce")
        if label_norm and (label_norm in row_norm or row_norm in label_norm):
            score += 5
        if pd.notna(year_match) and pd.notna(row_year) and int(year_match) == int(row_year):
            score += 4
        if pd.notna(coupon_num) and pd.notna(row_coupon) and abs(float(coupon_num) - float(row_coupon)) <= 0.05:
            score += 4
        if "convert" in row_norm and "note" in row_norm:
            score += 1
        if score > best_score:
            best_score = score
            best_row = row
    return best_row if best_score >= 4 else None


def _quarter_values_from_df(df: Any) -> List[str]:
    if df is None or getattr(df, "empty", True) or "quarter" not in df.columns:
        return []
    q = pd.to_datetime(df["quarter"], errors="coerce").dropna()
    return sorted({pd.Timestamp(v).date().isoformat() for v in q})


def _visible_sheet_names_from_path(path: Path) -> List[str]:
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        return [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    finally:
        wb.close()


def _ws_value(ws: Any, row: int, col: Optional[int]) -> Any:
    if not col:
        return None
    return ws.cell(row, col).value


def _dateish_to_iso(value: Any) -> str:
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        text = str(value or "").strip()
        return text
    return pd.Timestamp(ts).date().isoformat()


def _md_cell(value: Any) -> str:
    text = str(value or "").replace("\n", " ").replace("|", "\\|").strip()
    return text or "-"
