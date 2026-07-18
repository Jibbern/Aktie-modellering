"""Declarative workbook-module and profile contracts for new-ticker shells."""
from __future__ import annotations

import copy
import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from pbi_xbrl.json_schema_validation import load_json_strict, validate_json_schema
from pbi_xbrl.standard_template_formula_contract import formula_target_contracts


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MODULE_MANIFEST = ROOT / "docs" / "workbook_module_manifest.json"
MODULE_MANIFEST_SCHEMA = ROOT / "docs" / "workbook_module_manifest.schema.json"
MODULE_MANIFEST_VERSION = "1.0.0"
MODULE_CONTRACT_VERSION = "1.1.0"


@dataclass(frozen=True)
class ResolvedModuleProfile:
    profile_id: str
    enabled_modules: tuple[str, ...]
    profile_pack_ids: tuple[str, ...]
    dimensions: tuple[dict[str, str], ...]
    visible_sheet_order: tuple[str, ...]
    sheet_states: dict[str, str]

    def to_dict(self) -> dict[str, Any]:
        return {
            "profile_id": self.profile_id,
            "enabled_modules": list(self.enabled_modules),
            "profile_pack_ids": list(self.profile_pack_ids),
            "dimensions": [dict(row) for row in self.dimensions],
            "visible_sheet_order": list(self.visible_sheet_order),
            "sheet_states": dict(self.sheet_states),
        }


@dataclass(frozen=True)
class OwnedRange:
    owner_id: str
    contract_id: str
    sheet: str
    target: str
    kind: str
    exclusive_slot_id: str = ""
    source_id: str = ""


def load_workbook_module_manifest(path: Path | str = DEFAULT_MODULE_MANIFEST) -> dict[str, Any]:
    payload = load_json_strict(Path(path))
    if not isinstance(payload, dict):
        raise ValueError("Workbook module manifest must be a JSON object.")
    failures = validate_json_schema(payload, load_json_strict(MODULE_MANIFEST_SCHEMA))
    if failures:
        sample = "; ".join(f"{field} {keyword}: {message}" for field, keyword, message in failures[:10])
        raise ValueError(f"Workbook module manifest does not satisfy its schema: {sample}")
    semantic = validate_workbook_module_manifest(payload)
    if semantic:
        raise ValueError("Invalid workbook module manifest: " + "; ".join(semantic[:10]))
    return payload


def validate_workbook_module_manifest(payload: Mapping[str, Any]) -> list[str]:
    issues: list[str] = []
    modules = [row for row in payload.get("modules") or [] if isinstance(row, Mapping)]
    profiles = [row for row in payload.get("profiles") or [] if isinstance(row, Mapping)]
    packs = [row for row in payload.get("profile_packs") or [] if isinstance(row, Mapping)]
    module_ids = [str(row.get("module_id") or "") for row in modules]
    profile_ids = [str(row.get("profile_id") or "") for row in profiles]
    pack_ids = [str(row.get("pack_id") or "") for row in packs]
    issues.extend(_duplicate_issues(module_ids, "module_id"))
    issues.extend(_duplicate_issues(profile_ids, "profile_id"))
    issues.extend(_duplicate_issues(pack_ids, "pack_id"))
    module_set = set(module_ids)
    pack_set = set(pack_ids)
    module_by_id = {str(row.get("module_id") or ""): row for row in modules}
    for module in modules:
        module_id = str(module.get("module_id") or "")
        for dependency in module.get("dependencies") or []:
            if str(dependency) not in module_set:
                issues.append(f"Module {module_id!r} depends on unknown module {dependency!r}.")
    issues.extend(_dependency_cycle_issues(modules))

    sheet_contracts: dict[str, str] = {}
    binding_owners: dict[str, str] = {}
    for module in modules:
        module_id = str(module.get("module_id") or "")
        for sheet in module.get("sheets") or []:
            sheet_name = str(sheet.get("sheet") or "")
            if sheet_name in sheet_contracts:
                issues.append(
                    f"Sheet {sheet_name!r} is owned by both {sheet_contracts[sheet_name]!r} and {module_id!r}."
                )
            else:
                sheet_contracts[sheet_name] = module_id
            if str(sheet.get("role") or "") != "visible_product" and not list(sheet.get("headers") or []):
                issues.append(f"Hidden/module-capacity sheet {sheet_name!r} requires neutral headers.")
            if str(sheet.get("data_surface") or "binding_rows") == "formula_output":
                if str(sheet.get("role") or "") == "visible_product":
                    issues.append(f"Formula-output sheet {sheet_name!r} must be a hidden support surface.")
                if not str(sheet.get("formula_owner") or ""):
                    issues.append(f"Formula-output sheet {sheet_name!r} requires a formula_owner.")
        for binding_id in module.get("binding_ids") or []:
            binding_id = str(binding_id)
            if binding_id in binding_owners:
                issues.append(
                    f"Binding {binding_id!r} is owned by both {binding_owners[binding_id]!r} and {module_id!r}."
                )
            else:
                binding_owners[binding_id] = module_id

    union_order = [str(value) for value in payload.get("union_sheet_order") or []]
    if set(union_order) != set(sheet_contracts):
        missing = sorted(set(sheet_contracts) - set(union_order))
        unknown = sorted(set(union_order) - set(sheet_contracts))
        issues.append(f"union_sheet_order mismatch; missing={missing!r} unknown={unknown!r}.")

    blocks = visible_block_contracts(payload)
    block_ids = [row.contract_id for row in blocks]
    issues.extend(_duplicate_issues(block_ids, "visible block_id"))
    for block in blocks:
        if block.sheet not in sheet_contracts:
            issues.append(f"Visible block {block.contract_id!r} maps to unknown sheet {block.sheet!r}.")
            continue
        sheet_owner = sheet_contracts[block.sheet]
        if sheet_owner == block.owner_id:
            continue
        dependencies = _transitive_dependencies(block.owner_id, module_by_id)
        if sheet_owner not in dependencies:
            issues.append(
                f"Module {block.owner_id!r} owns visible block {block.contract_id!r} on sheet "
                f"{block.sheet!r}, owned by {sheet_owner!r}, without a direct or transitive dependency."
            )
    for index, left in enumerate(blocks):
        for right in blocks[index + 1 :]:
            if left.sheet != right.sheet or not _ranges_overlap(left.target, right.target):
                continue
            explicitly_exclusive = bool(
                left.exclusive_slot_id
                and left.exclusive_slot_id == right.exclusive_slot_id
                and left.kind == right.kind == "profile_pack_block"
            )
            if not explicitly_exclusive:
                issues.append(
                    f"Visible blocks {left.contract_id!r} ({left.owner_id!r}) and "
                    f"{right.contract_id!r} ({right.owner_id!r}) overlap on {left.sheet!r}."
                )

    styles = style_range_contracts(payload)
    style_ids = [row.contract_id for row in styles]
    issues.extend(_duplicate_issues(style_ids, "style_id"))
    style_keys = {(row.owner_id, row.sheet, row.target) for row in styles}
    block_keys = {(row.owner_id, row.sheet, row.target) for row in blocks}
    if style_keys != block_keys:
        missing = sorted(block_keys - style_keys)
        unknown = sorted(style_keys - block_keys)
        issues.append(f"Style-range ownership mismatch; missing={missing!r} unknown={unknown!r}.")

    declared_formula_ids = [
        str(formula_id)
        for module in modules
        for formula_id in module.get("formula_ids") or []
    ]
    issues.extend(_duplicate_issues(declared_formula_ids, "formula_id"))
    executable_formula_ids = {contract.formula_id for contract in formula_target_contracts()}
    declared_formula_set = set(declared_formula_ids)
    if declared_formula_set != executable_formula_ids:
        missing = sorted(executable_formula_ids - declared_formula_set)
        unknown = sorted(declared_formula_set - executable_formula_ids)
        issues.append(f"Formula ownership mismatch; missing={missing!r} unknown={unknown!r}.")
    formula_owner_map = formula_owners(payload)
    for contract in formula_target_contracts():
        owner = formula_owner_map.get(contract.formula_id, "")
        if not owner:
            continue
        for target in contract.targets:
            if not _range_is_owned(payload, owner, contract.sheet, target):
                issues.append(
                    f"Formula {contract.formula_id!r} target {contract.sheet}!{target} is outside "
                    f"module {owner!r} ownership."
                )

    declared_names = [
        str(name)
        for module in modules
        for name in module.get("defined_name_ids") or []
    ]
    issues.extend(_duplicate_issues(declared_names, "defined_name_id"))

    dimension_contract_ids = [str(value) for value in payload.get("dimension_contract_ids") or []]
    issues.extend(_duplicate_issues(dimension_contract_ids, "dimension_contract_id"))
    dimension_contract_set = set(dimension_contract_ids)

    pack_by_id = {str(row.get("pack_id") or ""): row for row in packs}
    for pack in packs:
        pack_id = str(pack.get("pack_id") or "")
        host = str(pack.get("host_module_id") or "")
        if host not in module_set:
            issues.append(f"Profile pack {pack_id!r} has unknown host module {host!r}.")
        elif str(module_by_id[host].get("module_type") or "") != "profile_pack_host":
            issues.append(f"Profile pack {pack_id!r} host {host!r} is not a profile_pack_host module.")
        driver_ids = [str(value) for value in pack.get("scenario_driver_ids") or []]
        issues.extend(
            f"Profile pack {pack_id!r}: {issue}"
            for issue in _duplicate_issues(driver_ids, "scenario_driver_id")
        )

    profile_by_id = {str(row.get("profile_id") or ""): row for row in profiles}
    union_profile = str(payload.get("union_shell_profile_id") or "")
    if union_profile not in profile_by_id:
        issues.append(f"Union shell profile {union_profile!r} does not exist.")
    for profile in profiles:
        profile_id = str(profile.get("profile_id") or "")
        enabled = {str(value) for value in profile.get("enabled_modules") or []}
        unknown_modules = sorted(enabled - module_set)
        if unknown_modules:
            issues.append(f"Profile {profile_id!r} enables unknown modules {unknown_modules!r}.")
        unknown_packs = sorted({str(value) for value in profile.get("profile_pack_ids") or []} - pack_set)
        if unknown_packs:
            issues.append(f"Profile {profile_id!r} selects unknown profile packs {unknown_packs!r}.")
        selected_slot_ids: list[str] = []
        for pack_id in profile.get("profile_pack_ids") or []:
            pack = pack_by_id.get(str(pack_id))
            if not isinstance(pack, Mapping):
                continue
            host = str(pack.get("host_module_id") or "")
            if host not in enabled:
                issues.append(f"Profile {profile_id!r} selects pack {pack_id!r} without host module {host!r}.")
            selected_slot_ids.extend(
                str(block.get("exclusive_slot_id") or "")
                for block in pack.get("visible_blocks") or []
                if str(block.get("exclusive_slot_id") or "")
            )
        issues.extend(
            f"Profile {profile_id!r} selects multiple packs for exclusive slot {slot!r}."
            for slot in sorted({value for value in selected_slot_ids if selected_slot_ids.count(value) > 1})
        )
        dimensions = [row for row in profile.get("dimensions") or [] if isinstance(row, Mapping)]
        dimension_ids = [str(row.get("dimension_id") or "") for row in dimensions]
        issues.extend(
            f"Profile {profile_id!r}: {issue}"
            for issue in _duplicate_issues(dimension_ids, "dimension_id")
        )
        unknown_dimensions = sorted(set(dimension_ids) - dimension_contract_set)
        if unknown_dimensions:
            issues.append(f"Profile {profile_id!r} declares unknown dimensions {unknown_dimensions!r}.")
        for module_id in enabled:
            module = next((row for row in modules if row.get("module_id") == module_id), {})
            missing_dependencies = sorted({str(value) for value in module.get("dependencies") or []} - enabled)
            if missing_dependencies:
                issues.append(
                    f"Profile {profile_id!r} enables {module_id!r} without dependencies {missing_dependencies!r}."
                )

    ticker_map = payload.get("ticker_profile_map") if isinstance(payload.get("ticker_profile_map"), Mapping) else {}
    for ticker, profile_id in ticker_map.items():
        if str(profile_id) not in profile_by_id:
            issues.append(f"Ticker {ticker!r} maps to unknown profile {profile_id!r}.")

    legacy_rows = [row for row in payload.get("legacy_sheet_inventory") or [] if isinstance(row, Mapping)]
    legacy_names = [str(row.get("legacy_sheet") or "") for row in legacy_rows]
    issues.extend(_duplicate_issues(legacy_names, "legacy_sheet"))
    for row in legacy_rows:
        disposition = str(row.get("disposition") or "")
        legacy_class = str(row.get("legacy_class") or "")
        module_id = str(row.get("module_id") or "")
        union_sheet = str(row.get("union_sheet") or "")
        if disposition in {"union_shell", "fixture_capacity"}:
            if module_id not in module_set:
                issues.append(f"Legacy sheet {row.get('legacy_sheet')!r} has no valid module owner.")
            if union_sheet not in sheet_contracts:
                issues.append(f"Legacy sheet {row.get('legacy_sheet')!r} maps to unknown union sheet {union_sheet!r}.")
        if disposition == "external_detail" and not str(row.get("external_contract") or ""):
            issues.append(f"External legacy sheet {row.get('legacy_sheet')!r} lacks external_contract.")
        if disposition == "rejected_redundant" and not str(row.get("replacement") or ""):
            issues.append(f"Rejected legacy sheet {row.get('legacy_sheet')!r} lacks a replacement.")
        expected = {
            "A": "union_shell",
            "B": "union_shell",
            "C": "union_shell",
            "D": "external_detail",
            "E": "fixture_capacity",
            "F": "rejected_redundant",
        }.get(legacy_class)
        if expected and disposition != expected:
            issues.append(
                f"Legacy sheet {row.get('legacy_sheet')!r} class {legacy_class!r} must use {expected!r}, got {disposition!r}."
            )
    return issues


def validate_binding_module_ownership(
    module_payload: Mapping[str, Any],
    binding_payload: Mapping[str, Any],
) -> list[str]:
    """Require one declared module owner for every executable binding."""

    owners = binding_owners(module_payload)
    binding_ids = [
        str(row.get("binding_id") or "")
        for row in binding_payload.get("bindings") or []
        if isinstance(row, Mapping)
    ]
    issues = _duplicate_issues(binding_ids, "binding_id")
    missing = sorted(set(binding_ids) - set(owners))
    unknown = sorted(set(owners) - set(binding_ids))
    if missing:
        issues.append(f"Bindings lack module ownership: {missing!r}.")
    if unknown:
        issues.append(f"Module manifest owns unknown bindings: {unknown!r}.")
    return issues


def visible_block_contracts(payload: Mapping[str, Any]) -> tuple[OwnedRange, ...]:
    rows: list[OwnedRange] = []
    for module in payload.get("modules") or []:
        owner = str(module.get("module_id") or "")
        for block in module.get("visible_blocks") or []:
            rows.append(
                OwnedRange(
                    owner,
                    str(block.get("block_id") or ""),
                    str(block.get("sheet") or ""),
                    str(block.get("target") or ""),
                    "module_block",
                    str(block.get("exclusive_slot_id") or ""),
                    owner,
                )
            )
    for pack in payload.get("profile_packs") or []:
        owner = str(pack.get("host_module_id") or "")
        pack_id = str(pack.get("pack_id") or "")
        for block in pack.get("visible_blocks") or []:
            rows.append(
                OwnedRange(
                    owner,
                    str(block.get("block_id") or ""),
                    str(block.get("sheet") or ""),
                    str(block.get("target") or ""),
                    "profile_pack_block",
                    str(block.get("exclusive_slot_id") or ""),
                    pack_id,
                )
            )
    return tuple(rows)


def style_range_contracts(payload: Mapping[str, Any]) -> tuple[OwnedRange, ...]:
    rows: list[OwnedRange] = []
    for module in payload.get("modules") or []:
        owner = str(module.get("module_id") or "")
        for style in module.get("style_ownership") or []:
            rows.append(
                OwnedRange(
                    owner,
                    str(style.get("style_id") or ""),
                    str(style.get("sheet") or ""),
                    str(style.get("target") or ""),
                    "style_range",
                    source_id=owner,
                )
            )
    return tuple(rows)


def formula_owners(payload: Mapping[str, Any]) -> dict[str, str]:
    return {
        str(formula_id): str(module.get("module_id") or "")
        for module in payload.get("modules") or []
        for formula_id in module.get("formula_ids") or []
    }


def defined_name_owners(payload: Mapping[str, Any]) -> dict[str, str]:
    return {
        str(name): str(module.get("module_id") or "")
        for module in payload.get("modules") or []
        for name in module.get("defined_name_ids") or []
    }


def enabled_formula_ids(payload: Mapping[str, Any], resolved: ResolvedModuleProfile) -> set[str]:
    enabled = set(resolved.enabled_modules)
    owners = formula_owners(payload)
    return {formula_id for formula_id, owner in owners.items() if owner in enabled}


def enabled_defined_name_ids(
    payload: Mapping[str, Any],
    binding_payload: Mapping[str, Any],
    resolved: ResolvedModuleProfile,
) -> set[str]:
    enabled = set(resolved.enabled_modules)
    names = {
        name
        for name, owner in defined_name_owners(payload).items()
        if owner in enabled
    }
    owners = binding_owners(payload)
    names.update(
        str(binding.get("binding_id") or "")
        for binding in binding_payload.get("bindings") or []
        if str(binding.get("planning_state") or "active") == "active"
        and owners.get(str(binding.get("binding_id") or "")) in enabled
    )
    return names


def validate_workbook_execution_ownership(
    workbook: Any,
    module_payload: Mapping[str, Any],
    binding_payload: Mapping[str, Any],
    resolved: ResolvedModuleProfile,
) -> list[str]:
    """Verify that the materialized workbook contains only profile-authorized formulas and names."""

    issues: list[str] = []
    formula_owner_map = formula_owners(module_payload)
    enabled = set(resolved.enabled_modules)
    cell_contracts: dict[tuple[str, str], str] = {}
    for contract in formula_target_contracts():
        for target in contract.targets:
            min_col, min_row, max_col, max_row = range_boundaries(target)
            for row in range(min_row, max_row + 1):
                for column in range(min_col, max_col + 1):
                    key = (contract.sheet, f"{get_column_letter(column)}{row}")
                    previous = cell_contracts.get(key)
                    if previous and previous != contract.formula_id:
                        issues.append(f"Formula cell {key[0]}!{key[1]} has contracts {previous!r} and {contract.formula_id!r}.")
                    cell_contracts[key] = contract.formula_id

    actual_formula_cells: set[tuple[str, str]] = set()
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    key = (ws.title, cell.coordinate)
                    actual_formula_cells.add(key)
                    formula_id = cell_contracts.get(key)
                    if not formula_id:
                        issues.append(f"Workbook formula {ws.title}!{cell.coordinate} has no module formula contract.")
                    elif formula_owner_map.get(formula_id) not in enabled:
                        issues.append(f"Disabled module formula {formula_id!r} remains at {ws.title}!{cell.coordinate}.")

    expected_formula_cells = {
        key
        for key, formula_id in cell_contracts.items()
        if formula_owner_map.get(formula_id) in enabled
    }
    missing_formulas = sorted(expected_formula_cells - actual_formula_cells)
    if missing_formulas:
        issues.append(f"Enabled formula cells are missing: {missing_formulas[:20]!r}.")

    expected_names = enabled_defined_name_ids(module_payload, binding_payload, resolved)
    actual_names = {str(name) for name in workbook.defined_names}
    missing_names = sorted(expected_names - actual_names)
    unknown_names = sorted(actual_names - expected_names)
    if missing_names:
        issues.append(f"Enabled defined names are missing: {missing_names[:20]!r}.")
    if unknown_names:
        issues.append(f"Workbook contains unauthorized defined names: {unknown_names[:20]!r}.")
    return issues


def resolve_module_profile(payload: Mapping[str, Any], profile_id: str) -> ResolvedModuleProfile:
    profile = next(
        (row for row in payload.get("profiles") or [] if str(row.get("profile_id") or "") == profile_id),
        None,
    )
    if not isinstance(profile, Mapping):
        raise ValueError(f"Unknown workbook module profile {profile_id!r}.")
    enabled = tuple(str(value) for value in profile.get("enabled_modules") or [])
    enabled_set = set(enabled)
    states: dict[str, str] = {}
    visible: list[str] = []
    contracts = sheet_contracts(payload)
    for sheet_name in payload.get("union_sheet_order") or []:
        sheet_name = str(sheet_name)
        sheet = contracts[sheet_name]
        module_id = str(sheet["module_id"])
        state = str(sheet["active_state"] if module_id in enabled_set else sheet["inactive_state"])
        states[sheet_name] = state
        if state == "visible":
            visible.append(sheet_name)
    if not visible:
        raise ValueError(f"Workbook module profile {profile_id!r} would create a workbook with no visible sheets.")
    return ResolvedModuleProfile(
        profile_id=profile_id,
        enabled_modules=enabled,
        profile_pack_ids=tuple(str(value) for value in profile.get("profile_pack_ids") or []),
        dimensions=tuple(dict(row) for row in profile.get("dimensions") or []),
        visible_sheet_order=tuple(visible),
        sheet_states=states,
    )


def profile_id_for_ticker(payload: Mapping[str, Any], ticker: str) -> str:
    ticker_map = payload.get("ticker_profile_map") if isinstance(payload.get("ticker_profile_map"), Mapping) else {}
    profile_id = str(ticker_map.get(str(ticker or "").upper()) or "")
    if not profile_id:
        raise ValueError(f"Ticker {ticker!r} has no declarative workbook module profile.")
    return profile_id


def sheet_contracts(payload: Mapping[str, Any]) -> dict[str, dict[str, Any]]:
    contracts: dict[str, dict[str, Any]] = {}
    for module in payload.get("modules") or []:
        module_id = str(module.get("module_id") or "")
        for row in module.get("sheets") or []:
            contract = dict(row)
            contract["module_id"] = module_id
            contracts[str(row["sheet"])] = contract
    return contracts


def binding_owners(payload: Mapping[str, Any]) -> dict[str, str]:
    owners: dict[str, str] = {}
    for module in payload.get("modules") or []:
        module_id = str(module.get("module_id") or "")
        for binding_id in module.get("binding_ids") or []:
            owners[str(binding_id)] = module_id
    return owners


def build_profile_binding_payload(
    binding_payload: Mapping[str, Any],
    module_payload: Mapping[str, Any],
    resolved: ResolvedModuleProfile,
) -> dict[str, Any]:
    _require_authoritative_resolution(module_payload, resolved)
    ownership_issues = validate_binding_module_ownership(module_payload, binding_payload)
    if ownership_issues:
        raise ValueError("Invalid binding/module ownership: " + "; ".join(ownership_issues))
    result = copy.deepcopy(dict(binding_payload))
    owners = binding_owners(module_payload)
    enabled = set(resolved.enabled_modules)
    resolved_bindings: list[dict[str, Any]] = []
    for raw in binding_payload.get("bindings") or []:
        binding = copy.deepcopy(dict(raw))
        binding_id = str(binding.get("binding_id") or "")
        owner = owners.get(binding_id)
        if not owner:
            raise ValueError(f"Binding {binding_id!r} has no module owner.")
        binding["module_id"] = owner
        if owner in enabled:
            resolved_bindings.append(binding)
    result["module_manifest_version"] = str(module_payload["version"])
    result["module_contract_version"] = str(module_payload["module_contract_version"])
    result["module_profile_id"] = resolved.profile_id
    result["enabled_modules"] = list(resolved.enabled_modules)
    result["profile_pack_ids"] = list(resolved.profile_pack_ids)
    result["scenario_profile_packs"] = scenario_profile_pack_contracts(module_payload, resolved)
    result["module_manifest_signature"] = canonical_json_sha256(module_payload)
    result["module_profile_signature"] = canonical_json_sha256(resolved.to_dict())
    result["bindings"] = resolved_bindings
    return result


def scenario_profile_pack_contracts(
    module_payload: Mapping[str, Any],
    resolved: ResolvedModuleProfile,
) -> list[dict[str, Any]]:
    """Project the exact pack-to-driver vocabulary allowed by a resolved profile."""

    _require_authoritative_resolution(module_payload, resolved)
    if resolved.profile_id == str(module_payload.get("union_shell_profile_id") or ""):
        selected = {str(pack.get("pack_id") or "") for pack in module_payload.get("profile_packs") or []}
    else:
        selected = set(resolved.profile_pack_ids)
    return [
        {
            "profile_pack_id": str(pack.get("pack_id") or ""),
            "scenario_driver_ids": sorted(str(driver_id) for driver_id in pack.get("scenario_driver_ids") or []),
        }
        for pack in module_payload.get("profile_packs") or []
        if isinstance(pack, Mapping) and str(pack.get("pack_id") or "") in selected
    ]


def build_profile_shell_manifest(
    manifest: Mapping[str, Any],
    module_payload: Mapping[str, Any],
    resolved: ResolvedModuleProfile,
) -> dict[str, Any]:
    _require_authoritative_resolution(module_payload, resolved)
    result = copy.deepcopy(dict(manifest))
    existing = {str(row["sheet"]): copy.deepcopy(dict(row)) for row in manifest.get("sheets") or []}
    contracts = sheet_contracts(module_payload)
    sheets: list[dict[str, Any]] = []
    for sheet_name in module_payload.get("union_sheet_order") or []:
        sheet_name = str(sheet_name)
        contract = contracts[sheet_name]
        if str(contract["role"]) == "visible_product":
            row = existing.get(sheet_name) or _reserved_sheet_manifest(contract)
        else:
            row = _project_hidden_support_manifest(contract, existing.get(sheet_name))
        row["module_id"] = str(contract["module_id"])
        row["module_role"] = str(contract["role"])
        row["legacy_class"] = str(contract["legacy_class"])
        row["state"] = resolved.sheet_states[sheet_name]
        sheets.append(row)
    result["visible_sheet_order"] = list(resolved.visible_sheet_order)
    result["union_sheet_order"] = [str(value) for value in module_payload.get("union_sheet_order") or []]
    result["sheets"] = sheets
    enabled_names = {
        name
        for name, owner in defined_name_owners(module_payload).items()
        if owner in set(resolved.enabled_modules)
    }
    result["required_anchors"] = [
        copy.deepcopy(dict(anchor))
        for anchor in manifest.get("required_anchors") or []
        if str(anchor.get("anchor_id") or "") in enabled_names
    ]
    result["module_manifest"] = {
        "path": "docs/workbook_module_manifest.json",
        "version": str(module_payload["version"]),
        "module_contract_version": str(module_payload["module_contract_version"]),
        "signature": canonical_json_sha256(module_payload),
    }
    result["module_profile"] = {
        **resolved.to_dict(),
        "enabled_formula_ids": sorted(enabled_formula_ids(module_payload, resolved)),
        "enabled_defined_name_ids": sorted(enabled_names),
        "signature": canonical_json_sha256(resolved.to_dict()),
    }
    result["external_detail_sheets"] = [
        str(row["legacy_sheet"])
        for row in module_payload.get("legacy_sheet_inventory") or []
        if row.get("disposition") == "external_detail"
    ]
    result["rejected_legacy_functions"] = [
        {"legacy_sheet": str(row["legacy_sheet"]), "replacement": str(row.get("replacement") or "")}
        for row in module_payload.get("legacy_sheet_inventory") or []
        if row.get("disposition") == "rejected_redundant"
    ]
    return result


def canonical_json_sha256(value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _reserved_sheet_manifest(contract: Mapping[str, Any]) -> dict[str, Any]:
    headers = [str(value) for value in contract.get("headers") or []]
    max_column = int(contract.get("capacity_columns") or max(1, len(headers)))
    max_row = int(contract.get("capacity_rows") or 5000)
    last_column = get_column_letter(max_column)
    sheet_name = str(contract["sheet"])
    binding_ranges = [str(value) for value in contract.get("binding_owned_ranges") or []]
    formula_ranges = [str(value) for value in contract.get("formula_owned_ranges") or []]
    reserved_ranges = [str(value) for value in contract.get("reserved_ranges") or []]
    sheet_protection = bool(contract.get("worksheet_protection"))
    if binding_ranges or formula_ranges or reserved_ranges:
        module_token = str(contract["module_id"])
        sheet_token = _safe_id(sheet_name)
        writable_zones = [
            {
                "zone_id": f"module_{module_token}_{sheet_token}_rows" if index == 0 else f"module_{module_token}_{sheet_token}_rows_{index + 1}",
                "target": target,
                "anchor_label": headers[0] if headers else "field",
                "value_shapes": ["table_rows"],
            }
            for index, target in enumerate(binding_ranges)
        ]
        non_writable_zones = [
            {
                "zone_id": f"module_{module_token}_{sheet_token}_headers",
                "target": f"A1:{last_column}1",
                "reason": "Neutral module support headers owned by the frozen shell.",
            },
            *[
                {
                    "zone_id": f"module_{module_token}_{sheet_token}_formula_{index + 1}",
                    "target": target,
                    "reason": "Formula-owned Hidden Value recomputation surface.",
                }
                for index, target in enumerate(formula_ranges)
            ],
            *[
                {
                    "zone_id": f"module_{module_token}_{sheet_token}_reserved_{index + 1}",
                    "target": target,
                    "reason": "Reserved neutral support capacity; no binding or formula ownership.",
                }
                for index, target in enumerate(reserved_ranges)
            ],
        ]
        return {
            "sheet": sheet_name,
            "static_layout_owner": "frozen_template_shell",
            "writable_zones": writable_zones,
            "non_writable_zones": non_writable_zones,
            "formulas_static_labels": headers,
            "worksheet_protection": sheet_protection,
        }
    if str(contract.get("data_surface") or "binding_rows") == "formula_output":
        return {
            "sheet": sheet_name,
            "static_layout_owner": "frozen_template_shell",
            "writable_zones": [],
            "non_writable_zones": [
                {
                    "zone_id": f"module_{str(contract['module_id'])}_{_safe_id(sheet_name)}_formula_surface",
                    "target": f"A1:{last_column}{max_row}",
                    "reason": "Formula-owned support surface; values are never written by bindings.",
                }
            ],
            "formulas_static_labels": headers,
            "worksheet_protection": sheet_protection,
        }
    return {
        "sheet": sheet_name,
        "static_layout_owner": "frozen_template_shell",
        "writable_zones": [
            {
                "zone_id": f"module_{str(contract['module_id'])}_{_safe_id(sheet_name)}_rows",
                "target": f"A2:{last_column}{max_row}",
                "anchor_label": headers[0] if headers else "field",
                "value_shapes": ["table_rows"],
            }
        ],
        "non_writable_zones": [
            {
                "zone_id": f"module_{str(contract['module_id'])}_{_safe_id(sheet_name)}_headers",
                "target": f"A1:{last_column}1",
                "reason": "Neutral module support headers owned by the frozen shell.",
            }
        ],
        "formulas_static_labels": headers,
        "worksheet_protection": sheet_protection,
    }


def _project_hidden_support_manifest(
    contract: Mapping[str, Any],
    existing: Mapping[str, Any] | None,
) -> dict[str, Any]:
    """Refresh support capacity while preserving established shell-zone IDs."""

    projected = _reserved_sheet_manifest(contract)
    if not existing:
        return projected
    for zone_key in ("writable_zones", "non_writable_zones"):
        old_zones = [row for row in existing.get(zone_key) or [] if isinstance(row, Mapping)]
        new_zones = projected[zone_key]
        if len(old_zones) == len(new_zones) == 1 and str(old_zones[0].get("zone_id") or ""):
            new_zones[0]["zone_id"] = str(old_zones[0]["zone_id"])
    return projected


def _dependency_cycle_issues(modules: Sequence[Mapping[str, Any]]) -> list[str]:
    graph = {
        str(row.get("module_id") or ""): tuple(str(value) for value in row.get("dependencies") or [])
        for row in modules
    }
    issues: list[str] = []
    visiting: set[str] = set()
    visited: set[str] = set()

    def visit(module_id: str, path: tuple[str, ...]) -> None:
        if module_id in visiting:
            issues.append("Module dependency cycle: " + " -> ".join((*path, module_id)))
            return
        if module_id in visited:
            return
        visiting.add(module_id)
        for dependency in graph.get(module_id, ()):
            visit(dependency, (*path, module_id))
        visiting.remove(module_id)
        visited.add(module_id)

    for module_id in graph:
        visit(module_id, tuple())
    return issues


def _transitive_dependencies(
    module_id: str,
    module_by_id: Mapping[str, Mapping[str, Any]],
) -> set[str]:
    dependencies: set[str] = set()
    pending = [str(value) for value in module_by_id.get(module_id, {}).get("dependencies") or []]
    while pending:
        dependency = pending.pop()
        if dependency in dependencies:
            continue
        dependencies.add(dependency)
        dependency_contract = module_by_id.get(dependency)
        if dependency_contract is not None:
            pending.extend(str(value) for value in dependency_contract.get("dependencies") or [])
    return dependencies


def _require_authoritative_resolution(payload: Mapping[str, Any], resolved: ResolvedModuleProfile) -> None:
    authoritative = resolve_module_profile(payload, resolved.profile_id)
    if authoritative.to_dict() != resolved.to_dict():
        raise ValueError(f"Resolved module profile {resolved.profile_id!r} no longer matches the manifest.")


def _ranges_overlap(left: str, right: str) -> bool:
    left_min_col, left_min_row, left_max_col, left_max_row = range_boundaries(left)
    right_min_col, right_min_row, right_max_col, right_max_row = range_boundaries(right)
    return not (
        left_max_col < right_min_col
        or right_max_col < left_min_col
        or left_max_row < right_min_row
        or right_max_row < left_min_row
    )


def _range_contains(container: str, target: str) -> bool:
    min_col, min_row, max_col, max_row = range_boundaries(container)
    target_min_col, target_min_row, target_max_col, target_max_row = range_boundaries(target)
    return (
        min_col <= target_min_col
        and min_row <= target_min_row
        and max_col >= target_max_col
        and max_row >= target_max_row
    )


def _range_is_owned(payload: Mapping[str, Any], owner: str, sheet: str, target: str) -> bool:
    if any(
        block.owner_id == owner
        and block.sheet == sheet
        and _range_contains(block.target, target)
        for block in visible_block_contracts(payload)
    ):
        return True
    for module in payload.get("modules") or []:
        if str(module.get("module_id") or "") != owner:
            continue
        for sheet_contract in module.get("sheets") or []:
            if str(sheet_contract.get("sheet") or "") != sheet:
                continue
            if str(sheet_contract.get("role") or "") == "visible_product":
                return False
            min_col, min_row, max_col, max_row = range_boundaries(target)
            return (
                min_col >= 1
                and min_row >= 1
                and max_col <= int(sheet_contract.get("capacity_columns") or 0)
                and max_row <= int(sheet_contract.get("capacity_rows") or 0)
            )
    return False


def _duplicate_issues(values: Sequence[str], label: str) -> list[str]:
    duplicates = sorted({value for value in values if values.count(value) > 1})
    return [f"Duplicate {label} {value!r}." for value in duplicates]


def _safe_id(value: str) -> str:
    return "".join(char.lower() if char.isalnum() else "_" for char in value).strip("_")
