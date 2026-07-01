"""Saved-workbook validation runner for delivered stock model workbooks.

The checks in this module are intentionally workbook-readback checks.  They
validate the files that a user opens, not only the in-memory writer objects.
"""
from __future__ import annotations

import argparse
import csv
import json
import posixpath
import re
import time
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from .path_config import resolve_stock_model_paths
from .workbook_quality_guardrails import run_workbook_quality_guardrails


TICKERS: Sequence[str] = ("PBI", "GPRE", "ANF")
ERROR_TOKENS: Sequence[str] = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")
REQUIRED_NAMED_RANGES: Sequence[str] = (
    "CompanyOperatingMargin_Latest",
    "OperatingMargin_Latest",
    "CompanyOperatingMargin_TTM",
)
BAD_MARKER_TERMS: Sequence[str] = (
    "[UPDATED]",
    "source_txt_file",
    "source_txt",
    "raw_json",
    "DEBUG",
    "TODO",
    "FIXME",
    "Actual / latest actual",
    "Adjusted EPS / EPS",
    "Found:",
    "Base active values",
    "Separate revenue cut; not summed",
)
GUIDANCE_NORMALIZED_MARKER_TICKERS = {"PBI", "GPRE"}
USER_FACING_SHEETS: Sequence[str] = (
    "SUMMARY",
    "Valuation",
    "{ticker}_Investment_Case",
    "Promise_Progress_UI",
    "Quarter_Notes_UI",
    "Operating_Drivers",
)
REQUIRED_SHARED_SHEETS: Sequence[str] = (
    "Valuation",
    "Promise_Progress_UI",
    "Quarter_Notes_UI",
    "History_Q",
    "Operating_Drivers",
    "Needs_Review",
    "QA_Log",
    "QA_Checks",
    "Scenario_Bridge_Tax_Treatment",
    "Scenario_Driver_Assumptions",
    "Quarter_Narrative_Data",
    "BS_Segments",
)
FULL_SCAN_SHEET_TEMPLATES: Sequence[str] = (
    "SUMMARY",
    "Valuation",
    "{ticker}_Investment_Case",
    "Promise_Progress_UI",
    "Quarter_Notes_UI",
    "Quarter_Narrative_Data",
    "Operating_Drivers",
    "Needs_Review",
    "QA_Log",
    "QA_Checks",
    "Scenario_Bridge_Tax_Treatment",
    "Scenario_Driver_Assumptions",
    "BS_Segments",
    "Guidance_Normalized",
    "Slides_Guidance",
)
TARGETED_SCAN_SHEETS: Sequence[str] = (
    "History_Q",
    "DATA_Facts_Long",
    "SEC_Audit_Log",
    "Quarter_Notes",
    "Quarter_Notes_Evidence",
    "Promise_Progress",
    "Promise_Evidence",
    "OCR_Text_Log",
    "Debt_Tranches_Q",
)
RAW_SHEET_NAME_HINTS: Sequence[str] = (
    "raw",
    "market",
    "basis_proxy",
)
CROSS_COMPANY_PATTERNS: Mapping[str, Sequence[str]] = {
    "PBI": (
        r"\bethanol\b",
        r"\b45Z\b",
        r"\bRINs?\b",
        r"\bRVO\b",
        r"\bcrush margin\b",
        r"\bAbercrombie\b",
        r"\bHollister\b",
    ),
    "GPRE": (
        r"\bPitney Bowes\b",
        r"\bPresort\b",
        r"\bSendTech\b",
        r"\bGEC\b",
        r"\bAbercrombie\b",
        r"\bHollister\b",
    ),
    "ANF": (
        r"\bPitney Bowes\b",
        r"\bPresort\b",
        r"\bSendTech\b",
        r"\bGEC\b",
        r"\bethanol\b",
        r"\b45Z\b",
        r"\bRINs?\b",
        r"\bRVO\b",
        r"\bcrush margin\b",
    ),
}
BAD_QUARTER_LABEL_PATTERNS: Sequence[str] = (
    r"\bQ[1-4]\s+FY\s*20\d{2}\b",
    r"\bQ[1-4]\s+20\d{2}\b",
    r"\bFY\s*20\d{2}\s+Q[1-4]\b",
    r"\b20\d{2}\s+Q[1-4]\b",
)
GOOD_QUARTER_LABEL_PATTERN = re.compile(r"\b20\d{2}-Q[1-4]\b")


@dataclass(frozen=True)
class ValidationIssue:
    category: str
    sheet: str = ""
    cell: str = ""
    value: str = ""
    detail: str = ""

    def to_dict(self) -> Dict[str, str]:
        return {
            "category": self.category,
            "sheet": self.sheet,
            "cell": self.cell,
            "value": self.value,
            "detail": self.detail,
        }


@dataclass(frozen=True)
class ValidationConfig:
    """Caps for expensive workbook readback scans.

    User-facing and QA sheets are still fully scanned. Large raw/source sheets
    are sampled so GPRE-style market exports do not dominate every validation
    run.
    """

    max_full_scan_rows: int = 10_000
    max_full_scan_cells: int = 250_000
    huge_sheet_row_threshold: int = 20_000
    sample_head_rows: int = 200
    sample_tail_rows: int = 200
    enable_quality_guardrails: bool = True
    quality_guardrails_warn_only: bool = False


@dataclass(frozen=True)
class SheetScanPlan:
    sheet_name: str
    mode: str
    max_row: int
    max_column: int
    reason: str = ""

    @property
    def sampled(self) -> bool:
        return self.mode in {"sampled", "large_sampled"}

    @property
    def large(self) -> bool:
        return self.mode == "large_sampled"


@dataclass
class WorkbookValidationResult:
    ticker: str
    path: str
    formula_error_count: int = 0
    needs_review_p1_count: int = 0
    qa_blank_nan_status_count: int = 0
    cross_company_leakage_count: int = 0
    bad_marker_count: int = 0
    quarter_label_issue_count: int = 0
    ooxml_table_issue_count: int = 0
    missing_required_sheets: List[str] = field(default_factory=list)
    missing_named_ranges: List[str] = field(default_factory=list)
    calc_settings_ok: bool = True
    quality_guardrail_p0_p1_count: int = 0
    quality_guardrail_p2_count: int = 0
    quality_guardrails_warn_only: bool = False
    quality_guardrail_issues: List[Dict[str, Any]] = field(default_factory=list)
    skipped_large_sheets: List[str] = field(default_factory=list)
    sampled_sheets: List[str] = field(default_factory=list)
    elapsed_seconds: float = 0.0
    category_elapsed_seconds: Dict[str, float] = field(default_factory=dict)
    issues: List[ValidationIssue] = field(default_factory=list)

    @property
    def required_sheets_ok(self) -> bool:
        return not self.missing_required_sheets

    @property
    def named_ranges_ok(self) -> bool:
        return not self.missing_named_ranges

    @property
    def overall(self) -> str:
        counters_ok = all(
            count == 0
            for count in [
                self.formula_error_count,
                self.needs_review_p1_count,
                self.qa_blank_nan_status_count,
                self.cross_company_leakage_count,
                self.bad_marker_count,
                self.quarter_label_issue_count,
                self.ooxml_table_issue_count,
                0 if self.quality_guardrails_warn_only else self.quality_guardrail_p0_p1_count,
            ]
        )
        return "PASS" if counters_ok and self.required_sheets_ok and self.named_ranges_ok and self.calc_settings_ok else "FAIL"

    def to_dict(self) -> Dict[str, Any]:
        return {
            "ticker": self.ticker,
            "path": self.path,
            "formula_errors": self.formula_error_count,
            "needs_review_p1": self.needs_review_p1_count,
            "qa_blank_nan": self.qa_blank_nan_status_count,
            "cross_company_leakage": self.cross_company_leakage_count,
            "bad_markers": self.bad_marker_count,
            "quarter_label_issues": self.quarter_label_issue_count,
            "ooxml_table_issues": self.ooxml_table_issue_count,
            "missing_required_sheets": self.missing_required_sheets,
            "missing_named_ranges": self.missing_named_ranges,
            "calc_settings_ok": self.calc_settings_ok,
            "quality_guardrail_p0_p1": self.quality_guardrail_p0_p1_count,
            "quality_guardrail_p2": self.quality_guardrail_p2_count,
            "quality_guardrails_warn_only": self.quality_guardrails_warn_only,
            "quality_guardrail_issues": self.quality_guardrail_issues,
            "skipped_large_sheets": self.skipped_large_sheets,
            "sampled_sheets": self.sampled_sheets,
            "elapsed_seconds": self.elapsed_seconds,
            "category_elapsed_seconds": self.category_elapsed_seconds,
            "overall": self.overall,
            "issues": [issue.to_dict() for issue in self.issues],
        }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _required_sheets_for_ticker(ticker: str) -> List[str]:
    return [*REQUIRED_SHARED_SHEETS, f"{ticker}_Investment_Case"]


def _user_facing_sheets_for_ticker(ticker: str, sheetnames: Sequence[str]) -> List[str]:
    out: List[str] = []
    for sheet in USER_FACING_SHEETS:
        resolved = sheet.format(ticker=ticker)
        if resolved in sheetnames:
            out.append(resolved)
    return out


def _resolve_sheet_templates(templates: Sequence[str], ticker: str, sheetnames: Sequence[str]) -> set[str]:
    out: set[str] = set()
    for template in templates:
        sheet_name = template.format(ticker=ticker)
        if sheet_name in sheetnames:
            out.add(sheet_name)
    return out


def _is_raw_or_market_sheet(sheet_name: str) -> bool:
    low = sheet_name.lower()
    if low == "economics_market_raw":
        return True
    if "guidance_raw" in low:
        return False
    return any(hint in low for hint in RAW_SHEET_NAME_HINTS) and not low.endswith("_ui")


def _sampled_row_numbers(max_row: int, *, head_rows: int, tail_rows: int) -> List[int]:
    if max_row <= 0:
        return []
    head_end = min(max_row, max(0, head_rows))
    tail_start = max(1, max_row - max(0, tail_rows) + 1)
    rows = set(range(1, head_end + 1))
    rows.update(range(tail_start, max_row + 1))
    return sorted(rows)


def _build_scan_plans(
    wb: Any,
    ticker: str,
    result: WorkbookValidationResult,
    config: ValidationConfig,
) -> Dict[str, SheetScanPlan]:
    full_scan_names = _resolve_sheet_templates(FULL_SCAN_SHEET_TEMPLATES, ticker, wb.sheetnames)
    plans: Dict[str, SheetScanPlan] = {}
    sampled: set[str] = set()
    skipped_large: set[str] = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            continue
        max_row = int(ws.max_row or 0)
        max_column = int(ws.max_column or 0)
        cell_count = max_row * max_column
        if sheet_name in full_scan_names or sheet_name.startswith("QA"):
            mode = "full"
            reason = "tier1_full"
        elif _is_raw_or_market_sheet(sheet_name) or max_row > config.huge_sheet_row_threshold:
            mode = "large_sampled"
            reason = "tier3_large_raw_or_market"
            skipped_large.add(sheet_name)
            sampled.add(sheet_name)
        elif (
            sheet_name in TARGETED_SCAN_SHEETS
            or max_row > config.max_full_scan_rows
            or cell_count > config.max_full_scan_cells
        ):
            mode = "sampled"
            reason = "tier2_targeted"
            sampled.add(sheet_name)
        else:
            mode = "full"
            reason = "small_sheet_full"
        plans[sheet_name] = SheetScanPlan(
            sheet_name=sheet_name,
            mode=mode,
            max_row=max_row,
            max_column=max_column,
            reason=reason,
        )
    result.sampled_sheets = sorted(sampled)
    result.skipped_large_sheets = sorted(skipped_large)
    return plans


def _append_issue(result: WorkbookValidationResult, issue: ValidationIssue) -> None:
    result.issues.append(issue)


def _iter_plan_cells(
    wb: Any,
    plans: Mapping[str, SheetScanPlan],
    config: ValidationConfig,
    sheet_names: Optional[Iterable[str]] = None,
) -> Iterable[Any]:
    names = list(sheet_names) if sheet_names is not None else list(wb.sheetnames)
    for sheet_name in names:
        plan = plans.get(sheet_name)
        if plan is None or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            continue
        if plan.mode == "full":
            row_numbers: Iterable[int] = range(1, plan.max_row + 1)
        else:
            row_numbers = _sampled_row_numbers(
                plan.max_row,
                head_rows=config.sample_head_rows,
                tail_rows=config.sample_tail_rows,
            )
        for rr in row_numbers:
            row = ws.iter_rows(
                min_row=rr,
                max_row=rr,
                min_col=1,
                max_col=max(1, plan.max_column),
            )
            for cell in row:
                yield from cell


def _scan_formula_errors(
    wb: Any,
    result: WorkbookValidationResult,
    plans: Mapping[str, SheetScanPlan],
    config: ValidationConfig,
) -> None:
    for cell in _iter_plan_cells(wb, plans, config):
        value = _cell_text(cell.value)
        if not value:
            continue
        for token in ERROR_TOKENS:
            if token in value:
                result.formula_error_count += 1
                _append_issue(
                    result,
                    ValidationIssue(
                        category="formula_error",
                        sheet=cell.parent.title,
                        cell=cell.coordinate,
                        value=value,
                        detail=f"{cell.parent.title}!{cell.coordinate} contains {token}: {value}",
                    ),
                )
                break


def _header_map(ws: Any) -> Dict[str, int]:
    return {
        _cell_text(ws.cell(1, cc).value).strip().lower(): cc
        for cc in range(1, int(ws.max_column or 0) + 1)
        if _cell_text(ws.cell(1, cc).value).strip()
    }


def _count_needs_review_p1(wb: Any, result: WorkbookValidationResult) -> None:
    if "Needs_Review" not in wb.sheetnames:
        return
    ws = wb["Needs_Review"]
    headers = _header_map(ws)
    cols = [
        col
        for header, col in headers.items()
        if "priority" in header or header in {"severity", "level"}
    ] or list(range(1, min(int(ws.max_column or 0), 8) + 1))
    for rr in range(2, int(ws.max_row or 0) + 1):
        for cc in cols:
            value = _cell_text(ws.cell(rr, cc).value).strip()
            if value.upper() == "P1":
                result.needs_review_p1_count += 1
                _append_issue(
                    result,
                    ValidationIssue(
                        category="needs_review_p1",
                        sheet=ws.title,
                        cell=ws.cell(rr, cc).coordinate,
                        value=value,
                        detail=f"{ws.title}!{ws.cell(rr, cc).coordinate} has P1 Needs_Review priority.",
                    ),
                )
                break


def _count_qa_blank_nan_status(wb: Any, result: WorkbookValidationResult) -> None:
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("QA"):
            continue
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            continue
        headers = _header_map(ws)
        status_cols = [col for header, col in headers.items() if "status" in header]
        if not status_cols:
            continue
        for rr in range(2, int(ws.max_row or 0) + 1):
            row_has_data = any(
                _cell_text(ws.cell(rr, cc).value).strip()
                for cc in range(1, min(int(ws.max_column or 0), 8) + 1)
            )
            if not row_has_data:
                continue
            for cc in status_cols:
                value = _cell_text(ws.cell(rr, cc).value).strip()
                if value.lower() in {"", "nan", "none", "null"}:
                    result.qa_blank_nan_status_count += 1
                    _append_issue(
                        result,
                        ValidationIssue(
                            category="qa_blank_nan_status",
                            sheet=ws.title,
                            cell=ws.cell(rr, cc).coordinate,
                            value=value,
                            detail=f"{ws.title}!{ws.cell(rr, cc).coordinate} has blank/nan QA status: {value!r}",
                        ),
                    )


def _scan_cross_company_leakage(
    wb: Any,
    ticker: str,
    result: WorkbookValidationResult,
    plans: Mapping[str, SheetScanPlan],
    config: ValidationConfig,
) -> None:
    patterns = [re.compile(pattern, flags=re.IGNORECASE) for pattern in CROSS_COMPANY_PATTERNS.get(ticker, ())]
    if not patterns:
        return
    for cell in _iter_plan_cells(wb, plans, config, _user_facing_sheets_for_ticker(ticker, wb.sheetnames)):
        value = _cell_text(cell.value)
        if not value:
            continue
        for pattern in patterns:
            if pattern.search(value):
                result.cross_company_leakage_count += 1
                _append_issue(
                    result,
                    ValidationIssue(
                        category="cross_company_leakage",
                        sheet=cell.parent.title,
                        cell=cell.coordinate,
                        value=value,
                        detail=f"{cell.parent.title}!{cell.coordinate} matches forbidden {pattern.pattern}: {value}",
                    ),
                )


def _bad_marker_terms_for_ticker(ticker: str) -> List[str]:
    terms = list(BAD_MARKER_TERMS)
    if ticker in GUIDANCE_NORMALIZED_MARKER_TICKERS:
        terms.append("Guidance_Normalized")
    return terms


def _scan_bad_markers(
    wb: Any,
    ticker: str,
    result: WorkbookValidationResult,
    plans: Mapping[str, SheetScanPlan],
    config: ValidationConfig,
) -> None:
    terms = _bad_marker_terms_for_ticker(ticker)
    for cell in _iter_plan_cells(wb, plans, config, _user_facing_sheets_for_ticker(ticker, wb.sheetnames)):
        value = _cell_text(cell.value)
        if not value:
            continue
        low = value.lower()
        for term in terms:
            if term.lower() in low:
                result.bad_marker_count += 1
                _append_issue(
                    result,
                    ValidationIssue(
                        category="bad_marker",
                        sheet=cell.parent.title,
                        cell=cell.coordinate,
                        value=value,
                        detail=f"{cell.parent.title}!{cell.coordinate} contains bad marker {term!r}: {value}",
                    ),
                )


def _scan_quarter_labels(
    wb: Any,
    ticker: str,
    result: WorkbookValidationResult,
    plans: Mapping[str, SheetScanPlan],
    config: ValidationConfig,
) -> None:
    patterns = [re.compile(pattern, flags=re.IGNORECASE) for pattern in BAD_QUARTER_LABEL_PATTERNS]
    for cell in _iter_plan_cells(wb, plans, config, _user_facing_sheets_for_ticker(ticker, wb.sheetnames)):
        value = _cell_text(cell.value)
        if not value:
            continue
        for pattern in patterns:
            if pattern.search(value):
                # Keep the rule tight: only flag if there is no accepted YYYY-QN
                # label in the same visible cell.
                if GOOD_QUARTER_LABEL_PATTERN.search(value):
                    continue
                result.quarter_label_issue_count += 1
                _append_issue(
                    result,
                    ValidationIssue(
                        category="quarter_label",
                        sheet=cell.parent.title,
                        cell=cell.coordinate,
                        value=value,
                        detail=f"{cell.parent.title}!{cell.coordinate} has non-standard quarter label: {value}",
                    ),
                )
                break


def _check_required_sheets(wb: Any, ticker: str, result: WorkbookValidationResult) -> None:
    for sheet in _required_sheets_for_ticker(ticker):
        if sheet not in wb.sheetnames:
            result.missing_required_sheets.append(sheet)
            _append_issue(
                result,
                ValidationIssue(
                    category="required_sheet",
                    sheet=sheet,
                    detail=f"Missing required sheet: {sheet}",
                ),
            )


def _check_named_ranges(wb: Any, result: WorkbookValidationResult) -> None:
    names = set(wb.defined_names.keys())
    for name in REQUIRED_NAMED_RANGES:
        if name not in names:
            result.missing_named_ranges.append(name)
            _append_issue(
                result,
                ValidationIssue(
                    category="named_range",
                    value=name,
                    detail=f"Missing expected named range: {name}",
                ),
            )


def _check_calc_settings(wb: Any, result: WorkbookValidationResult) -> None:
    calc = wb.calculation
    calc_mode = getattr(calc, "calcMode", None)
    full_calc = bool(getattr(calc, "fullCalcOnLoad", False))
    force_full = bool(getattr(calc, "forceFullCalc", False))
    result.calc_settings_ok = (calc_mode in {None, "auto"}) and full_calc and force_full
    if not result.calc_settings_ok:
        _append_issue(
            result,
            ValidationIssue(
                category="calc_settings",
                value=f"calcMode={calc_mode}; fullCalcOnLoad={full_calc}; forceFullCalc={force_full}",
                detail="Workbook calculation settings should be automatic with fullCalcOnLoad and forceFullCalc enabled.",
            ),
        )


_OOXML_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OOXML_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_OOXML_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_OOXML_NS = {
    "main": _OOXML_MAIN_NS,
    "rel": _OOXML_REL_NS,
}


def _ooxml_root(zf: ZipFile, part_name: str) -> ET.Element:
    return ET.fromstring(zf.read(part_name))


def _ooxml_rel_target_to_part(base_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), target))


def _ooxml_sheet_rels_part(sheet_part: str) -> str:
    return posixpath.join(posixpath.dirname(sheet_part), "_rels", posixpath.basename(sheet_part) + ".rels")


def _ooxml_cell_coordinate_col(cell_ref: str) -> int:
    letters = "".join(ch for ch in str(cell_ref or "") if ch.isalpha()).upper()
    col_idx = 0
    for ch in letters:
        col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
    return col_idx


def _ooxml_shared_strings(zf: ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = _ooxml_root(zf, "xl/sharedStrings.xml")
    values: List[str] = []
    for si in root.findall("main:si", _OOXML_NS):
        texts = [t.text or "" for t in si.findall(".//main:t", _OOXML_NS)]
        values.append("".join(texts))
    return values


def _ooxml_cell_value(cell: ET.Element, shared_strings: Sequence[str]) -> str:
    cell_type = cell.attrib.get("t")
    value = cell.find("main:v", _OOXML_NS)
    if cell_type == "s" and value is not None:
        try:
            return str(shared_strings[int(value.text or "0")])
        except (IndexError, ValueError):
            return ""
    if cell_type == "inlineStr":
        return "".join(t.text or "" for t in cell.findall(".//main:t", _OOXML_NS))
    return "" if value is None or value.text is None else str(value.text)


def _ooxml_header_values(
    zf: ZipFile,
    sheet_part: str,
    ref: str,
    shared_strings: Sequence[str],
) -> List[str]:
    min_col, min_row, max_col, _max_row = range_boundaries(ref)
    sheet_root = _ooxml_root(zf, sheet_part)
    values_by_col: Dict[int, str] = {}
    for row in sheet_root.findall(".//main:sheetData/main:row", _OOXML_NS):
        try:
            if int(row.attrib.get("r", "0")) != int(min_row):
                continue
        except ValueError:
            continue
        for cell in row.findall("main:c", _OOXML_NS):
            col_idx = _ooxml_cell_coordinate_col(cell.attrib.get("r", ""))
            if min_col <= col_idx <= max_col:
                values_by_col[col_idx] = _ooxml_cell_value(cell, shared_strings)
        break
    return [values_by_col.get(col_idx, "") for col_idx in range(min_col, max_col + 1)]


def _valid_excel_table_name(name: str) -> bool:
    if not name:
        return False
    if len(name) > 255:
        return False
    if re.search(r"\s", name):
        return False
    if not re.match(r"^[A-Za-z_\\]", name):
        return False
    if not re.match(r"^[A-Za-z_\\][A-Za-z0-9_.\\]*$", name):
        return False
    if re.match(r"^[A-Za-z]{1,3}[1-9][0-9]*$", name):
        return False
    return True


def _scan_ooxml_tables(workbook_path: Path, result: WorkbookValidationResult) -> None:
    """Validate OOXML table/header consistency that Excel repairs aggressively.

    This catches issues such as a table part declaring column names while the
    underlying worksheet header row is blank, a mismatched autoFilter/ref, or a
    tableColumns count that does not match the declared table range width.
    """

    try:
        with ZipFile(workbook_path) as zf:
            workbook_root = _ooxml_root(zf, "xl/workbook.xml")
            workbook_rels = _ooxml_root(zf, "xl/_rels/workbook.xml.rels")
            workbook_rel_targets = {
                rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
                for rel in workbook_rels.findall("rel:Relationship", _OOXML_NS)
            }
            sheet_parts: Dict[str, str] = {}
            sheets_el = workbook_root.find("main:sheets", _OOXML_NS)
            for sheet in list(sheets_el) if sheets_el is not None else []:
                sheet_name = sheet.attrib.get("name", "")
                rid = sheet.attrib.get(f"{{{_OOXML_OFFICE_REL_NS}}}id", "")
                target = workbook_rel_targets.get(rid, "")
                if sheet_name and target:
                    sheet_parts[sheet_name] = _ooxml_rel_target_to_part("xl/workbook.xml", target)

            table_to_sheet: Dict[str, str] = {}
            for sheet_name, sheet_part in sheet_parts.items():
                rels_part = _ooxml_sheet_rels_part(sheet_part)
                if rels_part not in zf.namelist():
                    continue
                rels_root = _ooxml_root(zf, rels_part)
                for rel in rels_root.findall("rel:Relationship", _OOXML_NS):
                    if str(rel.attrib.get("Type", "")).endswith("/table"):
                        table_part = _ooxml_rel_target_to_part(sheet_part, rel.attrib.get("Target", ""))
                        table_to_sheet[table_part] = sheet_name

            shared_strings = _ooxml_shared_strings(zf)
            table_names_seen: set[str] = set()
            for table_part in sorted(
                name for name in zf.namelist() if name.startswith("xl/tables/table") and name.endswith(".xml")
            ):
                table_root = _ooxml_root(zf, table_part)
                table_ref = str(table_root.attrib.get("ref") or "")
                table_name = str(table_root.attrib.get("name") or "")
                display_name = str(table_root.attrib.get("displayName") or "")
                sheet_name = table_to_sheet.get(table_part, "")
                columns_el = table_root.find("main:tableColumns", _OOXML_NS)
                table_columns = [str(col.attrib.get("name") or "") for col in list(columns_el) if columns_el is not None]
                declared_count = int(columns_el.attrib.get("count", "0")) if columns_el is not None else 0
                auto_filter_el = table_root.find("main:autoFilter", _OOXML_NS)
                auto_filter_ref = str(auto_filter_el.attrib.get("ref") or "") if auto_filter_el is not None else ""

                def add_issue(kind: str, detail: str) -> None:
                    result.ooxml_table_issue_count += 1
                    _append_issue(
                        result,
                        ValidationIssue(
                            category="ooxml_table",
                            sheet=sheet_name,
                            value=f"/{table_part}",
                            detail=f"{kind}: {detail}",
                        ),
                    )

                if not sheet_name:
                    add_issue("missing_sheet_relationship", f"Table part /{table_part} is not attached to a worksheet.")
                    continue
                if not table_ref:
                    add_issue("missing_ref", f"{display_name or table_name} has no table ref.")
                    continue
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(table_ref)
                except ValueError:
                    add_issue("invalid_ref", f"{display_name or table_name} has invalid ref {table_ref!r}.")
                    continue
                if max_row < min_row or max_col < min_col:
                    add_issue("empty_ref", f"{display_name or table_name} has empty ref {table_ref!r}.")
                    continue
                range_width = max_col - min_col + 1
                header_values = _ooxml_header_values(zf, sheet_parts[sheet_name], table_ref, shared_strings)
                header_text = [str(value or "").strip() for value in header_values]
                column_text = [str(value or "").strip() for value in table_columns]

                if auto_filter_ref and auto_filter_ref != table_ref:
                    add_issue(
                        "autofilter_ref_mismatch",
                        f"{sheet_name} {display_name or table_name} ref={table_ref} autoFilter={auto_filter_ref}.",
                    )
                if declared_count != range_width or len(table_columns) != range_width:
                    add_issue(
                        "table_columns_count_mismatch",
                        f"{sheet_name} {display_name or table_name} ref={table_ref} width={range_width} "
                        f"declared={declared_count} actual={len(table_columns)}.",
                    )
                if any(not value for value in column_text):
                    add_issue(
                        "blank_table_column_name",
                        f"{sheet_name} {display_name or table_name} has blank tableColumn names.",
                    )
                if len(column_text) != len(set(column_text)):
                    add_issue(
                        "duplicate_table_column_name",
                        f"{sheet_name} {display_name or table_name} has duplicate tableColumn names.",
                    )
                if any(not value for value in header_text):
                    add_issue(
                        "blank_worksheet_header_cell",
                        f"{sheet_name} {display_name or table_name} {table_ref} has blank worksheet header cells.",
                    )
                if len(header_text) != len(set(header_text)):
                    add_issue(
                        "duplicate_worksheet_header_cell",
                        f"{sheet_name} {display_name or table_name} {table_ref} has duplicate worksheet headers.",
                    )
                if header_text != column_text:
                    add_issue(
                        "worksheet_header_table_column_mismatch",
                        f"{sheet_name} {display_name or table_name} worksheet headers do not match tableColumns.",
                    )
                for candidate_name in [table_name, display_name]:
                    if not _valid_excel_table_name(candidate_name):
                        add_issue(
                            "invalid_table_name",
                            f"{sheet_name} table name/displayName {candidate_name!r} is not Excel-table-safe.",
                        )
                identity_name = display_name or table_name
                if identity_name in table_names_seen:
                    add_issue("duplicate_table_name", f"Duplicate table name/displayName: {identity_name}")
                if identity_name:
                    table_names_seen.add(identity_name)
    except (BadZipFile, KeyError, ET.ParseError, OSError, ValueError) as exc:
        result.ooxml_table_issue_count += 1
        _append_issue(
            result,
            ValidationIssue(
                category="ooxml_table",
                detail=f"Could not validate OOXML table package: {exc!r}",
            ),
        )


def _guardrail_failure_area(rule_id: str, owner: str, reason: str) -> str:
    blob = f"{rule_id} {owner} {reason}".lower()
    if "comparison_coloring" in blob or "color" in blob or "valuation hidden-value display" in blob:
        return "visual/style"
    if "amount extraction" in blob or "manual" in blob or "exception" in blob or "no-source" in blob or "annual-only" in blob:
        return "intentional exception missing"
    if "source" in blob or "coverage" in blob or "hydrate" in blob or "amount extraction" in blob:
        return "source coverage"
    if any(token in blob for token in ("exception", "manual", "not_applicable", "source_missing", "definition_mismatch", "annual_only")):
        return "intentional exception missing"
    return "model correctness"


def _quality_guardrail_payload(issue: Any, result: WorkbookValidationResult) -> Dict[str, Any]:
    payload = issue.to_dict()
    payload.update(
        {
            "guardrail_name": issue.rule_id,
            "workbook_path": result.path,
            "failure_area": _guardrail_failure_area(issue.rule_id, issue.owner, issue.reason),
        }
    )
    return payload


def _run_quality_guardrails(wb: Any, ticker: str, result: WorkbookValidationResult) -> None:
    seen: set[Tuple[Any, ...]] = set()
    for issue in run_workbook_quality_guardrails(wb, ticker):
        dedupe_key = (
            issue.severity,
            issue.rule_id,
            issue.ticker,
            issue.sheet,
            issue.row,
            issue.metric_label,
            issue.reason,
        )
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        payload = _quality_guardrail_payload(issue, result)
        result.quality_guardrail_issues.append(payload)
        if issue.severity in {"P0", "P1"}:
            result.quality_guardrail_p0_p1_count += 1
        elif issue.severity == "P2":
            result.quality_guardrail_p2_count += 1
        _append_issue(
            result,
            ValidationIssue(
                category=f"quality_guardrail_{issue.severity.lower()}",
                sheet=issue.sheet,
                cell=str(issue.row or ""),
                value=issue.metric_label,
                detail=f"{issue.rule_id}: {issue.reason} Owner/fix area: {issue.owner}",
            ),
        )


def _record_elapsed(result: WorkbookValidationResult, category: str, started: float) -> None:
    result.category_elapsed_seconds[category] = time.perf_counter() - started


def validate_workbook(
    path: Path | str,
    ticker: str,
    *,
    config: Optional[ValidationConfig] = None,
) -> WorkbookValidationResult:
    cfg = config or ValidationConfig()
    ticker_txt = str(ticker or "").strip().upper()
    workbook_path = Path(path)
    result = WorkbookValidationResult(
        ticker=ticker_txt,
        path=str(workbook_path),
        quality_guardrails_warn_only=bool(cfg.quality_guardrails_warn_only),
    )
    workbook_started = time.perf_counter()
    if not workbook_path.exists():
        _append_issue(
            result,
            ValidationIssue(
                category="workbook_missing",
                value=str(workbook_path),
                detail=f"Workbook does not exist: {workbook_path}",
            ),
        )
        result.missing_required_sheets = list(_required_sheets_for_ticker(ticker_txt))
        result.elapsed_seconds = time.perf_counter() - workbook_started
        return result

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        started = time.perf_counter()
        plans = _build_scan_plans(wb, ticker_txt, result, cfg)
        _record_elapsed(result, "scan_plan", started)

        started = time.perf_counter()
        _check_required_sheets(wb, ticker_txt, result)
        _record_elapsed(result, "required_sheets", started)

        started = time.perf_counter()
        _scan_formula_errors(wb, result, plans, cfg)
        _record_elapsed(result, "formula_errors", started)

        started = time.perf_counter()
        _count_needs_review_p1(wb, result)
        _count_qa_blank_nan_status(wb, result)
        _record_elapsed(result, "needs_review_qa", started)

        started = time.perf_counter()
        _scan_cross_company_leakage(wb, ticker_txt, result, plans, cfg)
        _scan_bad_markers(wb, ticker_txt, result, plans, cfg)
        _scan_quarter_labels(wb, ticker_txt, result, plans, cfg)
        _record_elapsed(result, "user_facing_text", started)

        started = time.perf_counter()
        _check_named_ranges(wb, result)
        _check_calc_settings(wb, result)
        _record_elapsed(result, "workbook_metadata", started)

        started = time.perf_counter()
        _scan_ooxml_tables(workbook_path, result)
        _record_elapsed(result, "ooxml_tables", started)

        if cfg.enable_quality_guardrails:
            started = time.perf_counter()
            _run_quality_guardrails(wb, ticker_txt, result)
            _record_elapsed(result, "quality_guardrails", started)
    finally:
        wb.close()
        result.elapsed_seconds = time.perf_counter() - workbook_started
    return result


def validate_workbooks(
    paths_by_ticker: Mapping[str, Path | str],
    *,
    config: Optional[ValidationConfig] = None,
) -> List[WorkbookValidationResult]:
    results: List[WorkbookValidationResult] = []
    for ticker in TICKERS:
        if ticker in paths_by_ticker:
            results.append(validate_workbook(paths_by_ticker[ticker], ticker, config=config))
    for ticker, path in paths_by_ticker.items():
        ticker_txt = str(ticker).upper()
        if ticker_txt not in TICKERS:
            results.append(validate_workbook(path, ticker_txt, config=config))
    return results


def summary_rows(results: Sequence[WorkbookValidationResult]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for result in results:
        rows.append(
            {
                "Ticker": result.ticker,
                "Formula errors": result.formula_error_count,
                "Needs_Review P1": result.needs_review_p1_count,
                "QA blank/nan": result.qa_blank_nan_status_count,
                "Cross-company leakage": result.cross_company_leakage_count,
                "Bad markers": result.bad_marker_count,
                "OOXML table issues": result.ooxml_table_issue_count,
                "Required sheets": "pass" if result.required_sheets_ok else f"missing {len(result.missing_required_sheets)}",
                "Named ranges": "pass" if result.named_ranges_ok else f"missing {len(result.missing_named_ranges)}",
                "Calc flags": "pass" if result.calc_settings_ok else "fail",
                "Guardrail P0/P1": result.quality_guardrail_p0_p1_count,
                "Guardrail P2": result.quality_guardrail_p2_count,
                "Skipped large sheets": len(result.skipped_large_sheets),
                "Sampled sheets": len(result.sampled_sheets),
                "Elapsed seconds": f"{result.elapsed_seconds:.2f}",
                "Overall": result.overall,
            }
        )
    return rows


def format_summary_table(results: Sequence[WorkbookValidationResult]) -> str:
    rows = summary_rows(results)
    headers = [
        "Ticker",
        "Formula errors",
        "Needs_Review P1",
        "QA blank/nan",
        "Cross-company leakage",
        "Bad markers",
        "OOXML table issues",
        "Required sheets",
        "Named ranges",
        "Calc flags",
        "Guardrail P0/P1",
        "Guardrail P2",
        "Skipped large sheets",
        "Sampled sheets",
        "Elapsed seconds",
        "Overall",
    ]
    widths = {header: max(len(header), *(len(str(row[header])) for row in rows)) for header in headers}
    lines = [" | ".join(header.ljust(widths[header]) for header in headers)]
    lines.append("-+-".join("-" * widths[header] for header in headers))
    for row in rows:
        lines.append(" | ".join(str(row[header]).ljust(widths[header]) for header in headers))
    return "\n".join(lines)


def write_validation_reports(
    results: Sequence[WorkbookValidationResult],
    output_dir: Path | str,
) -> Dict[str, Path]:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = out_dir / "workbook_validation_report.json"
    csv_path = out_dir / "workbook_validation_summary.csv"
    guardrails_json_path = out_dir / "workbook_validation_guardrails.json"
    guardrails_csv_path = out_dir / "workbook_validation_guardrails.csv"
    json_path.write_text(
        json.dumps([result.to_dict() for result in results], indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    rows = summary_rows(results)
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()) if rows else [])
        if rows:
            writer.writeheader()
            writer.writerows(rows)
    guardrail_rows = [
        issue
        for result in results
        for issue in result.quality_guardrail_issues
    ]
    guardrails_json_path.write_text(json.dumps(guardrail_rows, indent=2, ensure_ascii=False), encoding="utf-8")
    guardrail_fields = [
        "guardrail_name",
        "severity",
        "ticker",
        "workbook_path",
        "sheet",
        "row",
        "metric_label",
        "reason",
        "owner",
        "failure_area",
        "rule_id",
    ]
    with guardrails_csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=guardrail_fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(guardrail_rows)
    return {
        "json": json_path,
        "csv": csv_path,
        "guardrails_json": guardrails_json_path,
        "guardrails_csv": guardrails_csv_path,
    }


_WORKBOOK_EXTENSIONS = (".xlsx", ".xlsm")


def _is_workbook_file_path(path: Path) -> bool:
    return path.suffix.lower() in _WORKBOOK_EXTENSIONS


def _preferred_workbook_path(root: Path, ticker: str) -> Path:
    ticker_upper = str(ticker).upper()
    candidates = [root / f"{ticker_upper}_model{suffix}" for suffix in _WORKBOOK_EXTENSIONS]
    existing = [candidate for candidate in candidates if candidate.exists()]
    if existing:
        return max(existing, key=lambda candidate: candidate.stat().st_mtime)
    return candidates[0]


def _ticker_for_explicit_workbook(path: Path, tickers: Sequence[str]) -> str:
    normalized_tickers = [str(ticker).upper() for ticker in tickers]
    if len(normalized_tickers) == 1:
        return normalized_tickers[0]
    stem = path.stem.upper()
    matches = [ticker for ticker in normalized_tickers if ticker in stem]
    if len(matches) == 1:
        return matches[0]
    raise ValueError(
        "Explicit workbook file validation requires one ticker or a filename containing "
        f"exactly one ticker from {', '.join(normalized_tickers)}: {path}"
    )


def default_workbook_paths(workbook_dir: Path | str) -> Dict[str, Path]:
    root = Path(workbook_dir)
    return {ticker: _preferred_workbook_path(root, ticker) for ticker in TICKERS}


def resolve_workbook_paths(
    *,
    data_root: Path | str | None = None,
    workbook_dir: Path | str | None = None,
    tickers: Sequence[str] = TICKERS,
) -> Dict[str, Path]:
    if workbook_dir is not None and str(workbook_dir).strip():
        candidate = Path(workbook_dir).expanduser().resolve()
        if _is_workbook_file_path(candidate):
            ticker = _ticker_for_explicit_workbook(candidate, tickers)
            return {ticker: candidate}
        root = candidate
    else:
        root = resolve_workbook_dir(data_root=data_root, workbook_dir=None)
    return {
        str(ticker).upper(): _preferred_workbook_path(root, str(ticker).upper())
        for ticker in tickers
    }


def resolve_workbook_dir(
    *,
    data_root: Path | str | None = None,
    workbook_dir: Path | str | None = None,
) -> Path:
    if workbook_dir is not None and str(workbook_dir).strip():
        return Path(workbook_dir).expanduser().resolve()
    paths = resolve_stock_model_paths(Path(__file__).resolve().parents[2], data_root)
    return paths.excel_output_dir


def resolve_output_dir(
    *,
    data_root: Path | str | None = None,
    output_dir: Path | str | None = None,
) -> Path:
    if output_dir is not None and str(output_dir).strip():
        return Path(output_dir).expanduser().resolve()
    paths = resolve_stock_model_paths(Path(__file__).resolve().parents[2], data_root)
    return paths.validation_reports_dir / "workbook_validation"


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Validate saved stock model workbooks.")
    parser.add_argument(
        "--workbook-dir",
        default=None,
        help=(
            "Directory containing model workbooks, or one explicit .xlsx/.xlsm workbook "
            "path to validate exactly."
        ),
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Directory where JSON/CSV validation reports are written.",
    )
    parser.add_argument("--data-root", default="", help="Portable StockModelData root.")
    parser.add_argument("--tickers", nargs="*", default=list(TICKERS), help="Tickers to validate.")
    parser.add_argument("--max-full-scan-rows", type=int, default=ValidationConfig.max_full_scan_rows)
    parser.add_argument("--max-full-scan-cells", type=int, default=ValidationConfig.max_full_scan_cells)
    parser.add_argument("--huge-sheet-row-threshold", type=int, default=ValidationConfig.huge_sheet_row_threshold)
    parser.add_argument("--sample-head-rows", type=int, default=ValidationConfig.sample_head_rows)
    parser.add_argument("--sample-tail-rows", type=int, default=ValidationConfig.sample_tail_rows)
    parser.add_argument(
        "--skip-guardrails",
        action="store_true",
        help="Disable workbook quality guardrails; intended only for debugging validator behavior.",
    )
    parser.add_argument(
        "--no-quality-guardrails",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--guardrails-warn-only",
        action="store_true",
        help="Report workbook quality guardrails without letting P0/P1 issues fail validation.",
    )
    args = parser.parse_args(argv)

    output_dir = resolve_output_dir(data_root=args.data_root, output_dir=args.output_dir)
    paths = resolve_workbook_paths(
        data_root=args.data_root,
        workbook_dir=args.workbook_dir,
        tickers=args.tickers,
    )
    config = ValidationConfig(
        max_full_scan_rows=args.max_full_scan_rows,
        max_full_scan_cells=args.max_full_scan_cells,
        huge_sheet_row_threshold=args.huge_sheet_row_threshold,
        sample_head_rows=args.sample_head_rows,
        sample_tail_rows=args.sample_tail_rows,
        enable_quality_guardrails=not (args.no_quality_guardrails or args.skip_guardrails),
        quality_guardrails_warn_only=bool(args.guardrails_warn_only),
    )
    results = validate_workbooks(paths, config=config)
    report_paths = write_validation_reports(results, output_dir)
    print(format_summary_table(results))
    print(f"\nJSON report: {report_paths['json']}")
    print(f"CSV report: {report_paths['csv']}")
    print(f"Guardrail JSON report: {report_paths['guardrails_json']}")
    print(f"Guardrail CSV report: {report_paths['guardrails_csv']}")
    return 0 if all(result.overall == "PASS" for result in results) else 1


if __name__ == "__main__":
    raise SystemExit(main())
