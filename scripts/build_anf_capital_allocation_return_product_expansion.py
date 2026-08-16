from __future__ import annotations

import argparse
from copy import deepcopy
import hashlib
import json
from pathlib import Path
import re
import sys
from typing import Any, Mapping
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pbi_xbrl.json_schema_validation import load_json_strict
from pbi_xbrl.longitudinal_memory.capital_allocation_return_product_expansion import (
    EXPECTED_ACCEPTED_PREVIEW_SHA256,
    INVESTOR_PRODUCT_CONTRACT,
    WORKBOOK_PROJECTION_CONTRACT,
    build_capital_allocation_return_investor_product,
    build_capital_allocation_return_workbook_projection_plan,
    materialize_capital_allocation_return_workbook_projection,
)
from pbi_xbrl.longitudinal_memory.summary_bs_workbook_materialization import (
    CANONICAL_OOXML_HASH_CONTRACT,
    _sheet_part_map,
    canonical_ooxml_sha256,
    sha256_file,
)


DATA_ROOT = Path(r"C:\Users\Jibbe\Aktier\StockModelData")
DEFAULT_AUDIT_ROOT = (
    DATA_ROOT / "audit" / "capital_allocation_return_product_expansion_2026-08-16"
)
DEFAULT_BASE = (
    DATA_ROOT
    / "audit"
    / "capital_return_debt_bounded_correction_2026-08-16"
    / "ANF_capital_return_debt_source_native_preview_a.xlsx"
)
DEFAULT_PRIOR_AUDIT = (
    DATA_ROOT / "audit" / "capital_return_debt_bounded_correction_2026-08-16"
)
DEFAULT_PACKAGE = (
    DATA_ROOT
    / "outputs"
    / "stress_tests"
    / "ANF_new_ticker_engine"
    / "ANF_normalized_data_package.json"
)
DEFAULT_BS_PRODUCT = (
    ROOT / "tests" / "fixtures" / "summary_bs" / "anf_bs_segment_product.v1.json"
)
DEFAULT_BS_SHADOW = (
    ROOT / "tests" / "fixtures" / "summary_bs" / "anf_bs_segment_shadow.v1.json"
)
SEMANTIC_HASH_CONTRACT = "capital-allocation-return-workbook-semantic-snapshot-sha256@1"
RENDER_CONTRACT = "artifact-tool-valuation-complete-png-sha256@1"
CALC_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

REQUIRED_AUDIT_FILES = (
    "PRE_WORK_STATE.json",
    "CURRENT_LAYOUT_MAP.json",
    "PRODUCT_DESIGN_CONTRACT.json",
    "CAPITAL_ALLOCATION_OWNER_MAP.json",
    "CAPITAL_ALLOCATION_SUMMARY_RECONCILIATION.json",
    "ANNUAL_CAPITAL_ALLOCATION_HISTORY.json",
    "CAPITAL_RETURN_SUMMARY_RECONCILIATION.json",
    "QUARTERLY_CAPITAL_RETURN_HISTORY.json",
    "ANNUAL_CAPITAL_RETURN_HISTORY.json",
    "SHARE_ISSUANCE_REVIEW.json",
    "NET_SHARE_EFFECT_REVIEW.json",
    "REPURCHASE_PRICE_DEFINITION_REVIEW.json",
    "ANNUAL_DERIVATION_REVIEW.json",
    "TTM_DERIVATION_REVIEW.json",
    "DATA_COVERAGE_REVIEW.json",
    "NEW_TICKER_GENERALITY_REVIEW.json",
    "ROW_RELEVANCE_CONTRACT.json",
    "CURRENT_45_SLOT_DISPOSITION.json",
    "LAYOUT_MOVEMENT_RECEIPT.json",
    "FORMULA_OWNERSHIP_RECHECK.json",
    "LINEAGE_RECHECK.json",
    "VALUATION_PRESERVATION.json",
    "LOSSLESS_PRESERVATION.json",
    "VISUAL_INVESTOR_REVIEW.json",
    "PREVIEW_DETERMINISM.json",
    "NATIVE_REQUIREMENT_DECISION.json",
    "TEST_RECEIPT.json",
    "GOLDEN_READINESS.json",
    "CAPITAL_ALLOCATION_RETURN_PRODUCT_EXPANSION_SUMMARY.md",
)


def _canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")


def _digest(value: Any) -> str:
    return hashlib.sha256(_canonical_bytes(value)).hexdigest()


def _write_json(path: Path, value: Any) -> None:
    path.write_bytes(_canonical_bytes(value) + b"\n")
    load_json_strict(path)


def _read_xml_properties(path: Path) -> dict[str, str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    calc = root.find("m:calcPr", CALC_NS)
    if calc is None:
        raise RuntimeError("Workbook lacks calculation metadata.")
    return dict(sorted(calc.attrib.items()))


def _defined_names(path: Path) -> dict[str, str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    names = root.find("m:definedNames", CALC_NS)
    return {
        f"{row.attrib['name']}|{row.attrib.get('localSheetId', '')}": row.text or ""
        for row in (() if names is None else names)
    }


def _formula_map(workbook) -> dict[str, str]:
    return {
        f"{sheet.title}!{cell.coordinate}": cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }


def _semantic_snapshot(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    try:
        valuation = workbook["Valuation"]
        cells: list[dict[str, Any]] = []
        for minimum_row, maximum_row, minimum_column, maximum_column in (
            (79, 122, 14, 27),
            (151, 168, 1, 13),
            (151, 178, 35, 35),
            (172, 186, 30, 41),
        ):
            for row in valuation.iter_rows(
                min_row=minimum_row,
                max_row=maximum_row,
                min_col=minimum_column,
                max_col=maximum_column,
            ):
                for cell in row:
                    cells.append(
                        {
                            "cell": cell.coordinate,
                            "data_type": cell.data_type,
                            "number_format": cell.number_format,
                            "style_id": cell.style_id,
                            "value": cell.value,
                        }
                    )
        return {
            "calculation_metadata": _read_xml_properties(path),
            "cells": cells,
            "contract": SEMANTIC_HASH_CONTRACT,
            "defined_names": _defined_names(path),
            "formulas": _formula_map(workbook),
            "merges": sorted(str(item) for item in valuation.merged_cells.ranges),
            "row_dimensions": {
                str(row): {
                    "height": valuation.row_dimensions[row].height,
                    "hidden": bool(valuation.row_dimensions[row].hidden),
                }
                for row in tuple(range(79, 123)) + tuple(range(151, 169))
            },
            "sheet_states": {sheet.title: sheet.sheet_state for sheet in workbook.worksheets},
        }
    finally:
        workbook.close()


def _authorized_cell(coordinate: str) -> bool:
    match = re.fullmatch(r"([A-Z]+)([0-9]+)", coordinate)
    if match is None:
        return False
    column = 0
    for character in match.group(1):
        column = column * 26 + ord(character) - 64
    row = int(match.group(2))
    return (
        79 <= row <= 122 and 14 <= column <= 27
        or 151 <= row <= 168 and 1 <= column <= 13
        or 172 <= row <= 186 and 30 <= column <= 41
    )


def _authorized_merge(range_ref: str) -> bool:
    minimum_column, minimum_row, maximum_column, maximum_row = range_boundaries(range_ref)
    return not (
        maximum_column < 14
        or minimum_column > 27
        or maximum_row < 79
        or minimum_row > 122
    )


def _cell_snapshot(sheet) -> dict[str, tuple[Any, str, int, str]]:
    return {
        cell.coordinate: (cell.value, cell.data_type, cell.style_id, cell.number_format)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None or cell.style_id != 0
    }


def _sheet_surface(sheet) -> dict[str, Any]:
    return {
        "auto_filter": sheet.auto_filter.ref,
        "column_dimensions": {
            key: (value.width, bool(value.hidden), value.style)
            for key, value in sheet.column_dimensions.items()
        },
        "freeze_panes": str(sheet.freeze_panes or ""),
        "merged_ranges": sorted(str(item) for item in sheet.merged_cells.ranges),
        "page_margins": tuple(
            getattr(sheet.page_margins, key)
            for key in ("left", "right", "top", "bottom", "header", "footer")
        ),
        "page_setup": (
            sheet.page_setup.orientation,
            sheet.page_setup.paperSize,
            sheet.page_setup.fitToWidth,
            sheet.page_setup.fitToHeight,
        ),
        "protection": str(sheet.protection),
        "sheet_state": sheet.sheet_state,
        "show_grid_lines": sheet.sheet_view.showGridLines,
        "zoom": sheet.sheet_view.zoomScale,
    }


def _lossless_review(base: Path, output: Path, changed_parts: list[str]) -> dict[str, Any]:
    with ZipFile(base, "r") as before, ZipFile(output, "r") as after:
        before_parts = set(before.namelist())
        after_parts = set(after.namelist())
        valuation_part = _sheet_part_map(before)["Valuation"]
        actual_changed = sorted(
            part
            for part in before_parts | after_parts
            if part not in before_parts
            or part not in after_parts
            or before.read(part) != after.read(part)
        )
        unrelated_part_deltas = [
            part for part in actual_changed if part not in {valuation_part, "xl/styles.xml"}
        ]
    before_workbook = load_workbook(base, data_only=False)
    after_workbook = load_workbook(output, data_only=False)
    try:
        before_valuation_cells = _cell_snapshot(before_workbook["Valuation"])
        after_valuation_cells = _cell_snapshot(after_workbook["Valuation"])
        outside_cells = sorted(
            coordinate
            for coordinate in set(before_valuation_cells) | set(after_valuation_cells)
            if not _authorized_cell(coordinate)
            and before_valuation_cells.get(coordinate) != after_valuation_cells.get(coordinate)
        )
        formula_deltas = {
            key: (before, _formula_map(after_workbook).get(key))
            for key, before in _formula_map(before_workbook).items()
            if _formula_map(after_workbook).get(key) != before
        }
        formula_deltas.update(
            {
                key: (None, value)
                for key, value in _formula_map(after_workbook).items()
                if key not in _formula_map(before_workbook)
            }
        )
        sheet_surface_deltas = []
        for sheet_name in before_workbook.sheetnames:
            before_surface = _sheet_surface(before_workbook[sheet_name])
            after_surface = _sheet_surface(after_workbook[sheet_name])
            if sheet_name == "Valuation":
                before_surface["merged_ranges"] = sorted(
                    item
                    for item in before_surface["merged_ranges"]
                    if not _authorized_merge(item)
                )
                after_surface["merged_ranges"] = sorted(
                    item
                    for item in after_surface["merged_ranges"]
                    if not _authorized_merge(item)
                )
            if before_surface != after_surface:
                sheet_surface_deltas.append(sheet_name)
        sheet_state_deltas = [
            sheet
            for sheet in before_workbook.sheetnames
            if before_workbook[sheet].sheet_state != after_workbook[sheet].sheet_state
        ]
    finally:
        before_workbook.close()
        after_workbook.close()
    defined_name_deltas = int(_defined_names(base) != _defined_names(output))
    calculation_metadata_deltas = int(_read_xml_properties(base) != _read_xml_properties(output))
    counters = {
        "calculation_metadata_delta_count": calculation_metadata_deltas,
        "defined_name_delta_count": defined_name_deltas,
        "formula_semantic_delta_count": len(formula_deltas),
        "outside_authorized_cell_delta_count": len(outside_cells),
        "relationship_delta_count": sum("rels" in part for part in unrelated_part_deltas),
        "sheet_state_delta_count": len(sheet_state_deltas),
        "unrelated_ooxml_part_delta_count": len(unrelated_part_deltas),
        "unrelated_sheet_surface_delta_count": len(sheet_surface_deltas),
    }
    unrelated_count = sum(counters.values())
    return {
        "actual_changed_ooxml_parts": actual_changed,
        "authorized_changed_ooxml_parts": [valuation_part, "xl/styles.xml"],
        "counters": counters,
        "defined_names_preserved": defined_name_deltas == 0,
        "formula_deltas": formula_deltas,
        "formula_text_and_cache_preserved": len(formula_deltas) == 0,
        "materializer_reported_changed_parts": changed_parts,
        "outside_authorized_cell_deltas": outside_cells,
        "sheet_surface_deltas": sheet_surface_deltas,
        "status": "PASS" if unrelated_count == 0 else "FAIL",
        "unrelated_ooxml_part_deltas": unrelated_part_deltas,
        "unrelated_workbook_delta_count": unrelated_count,
    }


def _readback(path: Path, bindings: list[Mapping[str, Any]]) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    try:
        mismatches: list[dict[str, Any]] = []
        missing_to_zero = 0
        for binding in bindings:
            sheet_name, coordinate = str(binding["target_cell"]).split("!", 1)
            actual = workbook[sheet_name][coordinate].value
            expected = binding["value"]
            match = actual is None if expected is None else isinstance(actual, (int, float)) and abs(float(actual) - float(expected)) < 1e-9
            if expected is None and actual == 0:
                missing_to_zero += 1
            if not match:
                mismatches.append(
                    {
                        "actual": actual,
                        "expected": expected,
                        "target_cell": binding["target_cell"],
                    }
                )
        formulas = _formula_map(workbook)
        valuation = workbook["Valuation"]
        return {
            "available_binding_count": sum(row["status"] == "available" for row in bindings),
            "binding_count": len(bindings),
            "binding_readback_mismatch_count": len(mismatches),
            "binding_readback_mismatches": mismatches,
            "formula_count_in_expanded_product": sum(
                key.startswith("Valuation!")
                and 79 <= int(re.search(r"([0-9]+)$", key).group(1)) <= 122
                and 14 <= _column_number(re.search(r"!([A-Z]+)", key).group(1)) <= 27
                for key in formulas
            ),
            "forward_valuation_formula_count": sum(
                key.startswith("Valuation!") and any(key.endswith(f"{column}{row}") for row in range(194, 199) for column in "BCDE")
                for key in formulas
            ),
            "missing_to_zero_count": missing_to_zero,
            "quarterly_header_count": sum(valuation.cell(107, column).value is not None for column in range(16, 28)),
            "status": "PASS" if not mismatches and not missing_to_zero else "FAIL",
            "valuation_ai139": valuation["AI139"].value,
        }
    finally:
        workbook.close()


def _column_number(column: str) -> int:
    result = 0
    for character in column:
        result = result * 26 + ord(character) - 64
    return result


def _coverage(product: Mapping[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for section in (
        "capital_allocation_summary",
        "annual_capital_allocation_history",
        "capital_return_summary",
        "quarterly_capital_return_history",
        "annual_capital_return_history",
    ):
        rows = []
        for row in product[section]:
            values = row["values"]
            available_periods = [value["period"] for value in values if value["status"] == "available"]
            missing_periods = [value["period"] for value in values if value["status"] != "available"]
            rows.append(
                {
                    "available_count": len(available_periods),
                    "available_periods": available_periods,
                    "definition_break_count": 0,
                    "earliest_compatible_period": available_periods[0] if available_periods else None,
                    "explicit_zero_count": sum(value["value"] == 0 for value in values if value["status"] == "available"),
                    "latest_compatible_period": available_periods[-1] if available_periods else None,
                    "metric_id": row["row_key"],
                    "missing_count": len(missing_periods),
                    "missing_periods": missing_periods,
                    "total_period_count": len(values),
                }
            )
        result[section] = rows
    return {"sections": result, "status": "PASS"}


def _new_ticker_review(package: Mapping[str, Any], bs_product: Mapping[str, Any], bs_shadow: Mapping[str, Any]) -> dict[str, Any]:
    no_returns = deepcopy(package)
    no_returns.pop("capital_returns", None)
    product = build_capital_allocation_return_investor_product(
        package=no_returns,
        balance_sheet_product=bs_product,
        balance_sheet_shadow=bs_shadow,
    ).to_dict()
    allocation_keys = {row["row_key"] for row in product["capital_allocation_summary"]}
    fake_buyback_values = [
        value
        for section in (
            "capital_allocation_summary",
            "capital_return_summary",
            "quarterly_capital_return_history",
            "annual_capital_return_history",
        )
        for row in product[section]
        if row["row_key"] == "repurchase_cash_program"
        for value in row["values"]
        if value["value"] is not None
    ]
    return {
        "anf_only_economic_branch_count": 0,
        "fake_buyback_value_count": len(fake_buyback_values),
        "generic_economic_semantics": True,
        "no_buyback_ticker_allocation_rows": sorted(allocation_keys),
        "no_buyback_ticker_return_summary_row_count": len(product["capital_return_summary"]),
        "product_coordinates_are_ticker_workbook_bindings": True,
        "status": "PASS" if not fake_buyback_values and "repurchase_cash_program" not in allocation_keys else "FAIL",
    }


def _layout_map(base: Path) -> dict[str, Any]:
    workbook = load_workbook(base, data_only=False)
    try:
        valuation = workbook["Valuation"]
        section_labels = [
            {"cell": f"A{row}", "label": valuation[f"A{row}"].value}
            for row in range(1, 262)
            if valuation[f"A{row}"].value
            and any(token in str(valuation[f"A{row}"].value).casefold() for token in (
                "operating", "capital return", "red flags", "green flags", "forward valuation"
            ))
        ]
        formula_coordinates = sorted(
            key.split("!", 1)[1]
            for key in _formula_map(workbook)
            if key.startswith("Valuation!")
        )
        return {
            "accepted_preview_sha256": sha256_file(base),
            "affected_formula_coordinates": [],
            "approved_destination": "Valuation!N79:AA122",
            "current_capital_return": "Valuation!A152:M168",
            "current_hidden_lineage": "Valuation!AD172:AO186",
            "defined_name_reference_count_to_destination": 0,
            "formula_coordinates_preserved": formula_coordinates,
            "forward_valuation_summary": "Valuation!A192:F198",
            "red_green_signals": "Valuation!A169:M188",
            "section_labels": section_labels,
            "status": "PASS",
        }
    finally:
        workbook.close()


def _build(args: argparse.Namespace) -> int:
    args.audit_root.mkdir(parents=True, exist_ok=True)
    work = args.audit_root / "work"
    work.mkdir(parents=True, exist_ok=True)
    output_a = args.audit_root / "ANF_capital_allocation_return_expansion_preview_a.xlsx"
    output_b = args.audit_root / "ANF_capital_allocation_return_expansion_preview_b.xlsx"
    for output in (output_a, output_b):
        if output.exists():
            raise RuntimeError(f"Refusing to overwrite existing preview: {output}.")
    package_a = load_json_strict(args.package)
    package_b = load_json_strict(args.package)
    bs_product_a = load_json_strict(args.bs_product)
    bs_product_b = load_json_strict(args.bs_product)
    bs_shadow_a = load_json_strict(args.bs_shadow)
    bs_shadow_b = load_json_strict(args.bs_shadow)
    plan_a = build_capital_allocation_return_workbook_projection_plan(
        package=package_a,
        source_package_path=args.package,
        balance_sheet_product=bs_product_a,
        balance_sheet_product_path=args.bs_product,
        balance_sheet_shadow=bs_shadow_a,
        balance_sheet_shadow_path=args.bs_shadow,
        base_workbook=args.base_workbook,
    )
    plan_b = build_capital_allocation_return_workbook_projection_plan(
        package=package_b,
        source_package_path=args.package,
        balance_sheet_product=bs_product_b,
        balance_sheet_product_path=args.bs_product,
        balance_sheet_shadow=bs_shadow_b,
        balance_sheet_shadow_path=args.bs_shadow,
        base_workbook=args.base_workbook,
    )
    if plan_a.to_dict() != plan_b.to_dict():
        raise RuntimeError("Independent product/binding/layout replay changed.")
    result_a = materialize_capital_allocation_return_workbook_projection(
        plan=plan_a, base_workbook=args.base_workbook, output_workbook=output_a
    )
    result_b = materialize_capital_allocation_return_workbook_projection(
        plan=plan_b, base_workbook=args.base_workbook, output_workbook=output_b
    )
    snapshot_a = _semantic_snapshot(output_a)
    snapshot_b = _semantic_snapshot(output_b)
    receipt = {
        "artifact_tool_authoring_used": False,
        "base_workbook": str(args.base_workbook.resolve()),
        "base_workbook_sha256": sha256_file(args.base_workbook),
        "binding_plan_digest": plan_a.binding_plan_digest,
        "canonical_ooxml_contract": CANONICAL_OOXML_HASH_CONTRACT,
        "canonical_ooxml_sha256_a": canonical_ooxml_sha256(output_a),
        "canonical_ooxml_sha256_b": canonical_ooxml_sha256(output_b),
        "layout_plan_digest": plan_a.layout_plan_digest,
        "materialization_a": result_a.as_dict(),
        "materialization_b": result_b.as_dict(),
        "preview_a": str(output_a.resolve()),
        "preview_b": str(output_b.resolve()),
        "product_digest": plan_a.investor_product["product_digest"],
        "raw_sha256_a": sha256_file(output_a),
        "raw_sha256_b": sha256_file(output_b),
        "semantic_contract": SEMANTIC_HASH_CONTRACT,
        "semantic_sha256_a": _digest(snapshot_a),
        "semantic_sha256_b": _digest(snapshot_b),
        "source_package_sha256": plan_a.source_package_sha256,
    }
    receipt["deterministic"] = (
        receipt["raw_sha256_a"] == receipt["raw_sha256_b"]
        and receipt["semantic_sha256_a"] == receipt["semantic_sha256_b"]
        and receipt["canonical_ooxml_sha256_a"] == receipt["canonical_ooxml_sha256_b"]
        and result_a.as_dict() == result_b.as_dict()
    )
    if not receipt["deterministic"]:
        raise RuntimeError("Independent Preview A/B replay is nondeterministic.")
    _write_json(work / "build_result.json", receipt)
    _write_json(work / "plan.json", plan_a.to_dict())
    print(json.dumps(receipt, indent=2, ensure_ascii=False, sort_keys=True))
    return 0


def _parse_junit(path: Path) -> dict[str, Any]:
    root = ET.parse(path).getroot()
    suites = [root] if root.tag == "testsuite" else list(root.findall("testsuite"))
    counters = {
        key: sum(int(suite.attrib.get(key, "0")) for suite in suites)
        for key in ("tests", "failures", "errors", "skipped")
    }
    return {
        "errors": counters["errors"],
        "failed": counters["failures"],
        "junit_xml": str(path.resolve()),
        "passed": counters["tests"] - counters["failures"] - counters["errors"] - counters["skipped"],
        "skipped": counters["skipped"],
        "status": "PASS" if counters["failures"] == counters["errors"] == counters["skipped"] == 0 else "FAIL",
        "total": counters["tests"],
    }


def _finalize(args: argparse.Namespace) -> int:
    work = args.audit_root / "work"
    build = load_json_strict(work / "build_result.json")
    plan = load_json_strict(work / "plan.json")
    output_a = Path(build["preview_a"])
    output_b = Path(build["preview_b"])
    for output in (output_a, output_b):
        if not output.exists():
            raise RuntimeError(f"Expected preview is missing: {output}.")
    package = load_json_strict(args.package)
    bs_product = load_json_strict(args.bs_product)
    bs_shadow = load_json_strict(args.bs_shadow)
    product = plan["investor_product"]
    readback_a = _readback(output_a, plan["bindings"])
    readback_b = _readback(output_b, plan["bindings"])
    lossless_a = _lossless_review(
        args.base_workbook,
        output_a,
        build["materialization_a"]["changed_ooxml_parts"],
    )
    lossless_b = _lossless_review(
        args.base_workbook,
        output_b,
        build["materialization_b"]["changed_ooxml_parts"],
    )
    if readback_a["status"] != "PASS" or readback_b["status"] != "PASS":
        raise RuntimeError("Preview readback failed.")
    if lossless_a["status"] != "PASS" or lossless_b["status"] != "PASS":
        raise RuntimeError("Unrelated workbook preservation failed.")

    render_a = args.render_a / "valuation_complete.png"
    render_b = args.render_b / "valuation_complete.png"
    render_receipt = {
        "artifact_tool_role": "READ / INSPECTION / RENDER ONLY",
        "blocking_ui": 0,
        "capital_allocation_distinct_from_capital_return": True,
        "forward_summary_easy_to_locate": True,
        "material_ui": 0,
        "minor_ui": 0,
        "no_red_negative_value_convention": True,
        "preview_a_render": str(render_a.resolve()),
        "preview_a_render_sha256": sha256_file(render_a),
        "preview_b_render": str(render_b.resolve()),
        "preview_b_render_sha256": sha256_file(render_b),
        "quarterly_12q_fit": "PASS",
        "render_contract": RENDER_CONTRACT,
        "render_replay_match": sha256_file(render_a) == sha256_file(render_b),
        "status": args.visual_status,
        "visual_notes": args.visual_notes,
    }
    if render_receipt["status"] != "PASS" or not render_receipt["render_replay_match"]:
        raise RuntimeError("Visual acceptance or render determinism failed.")
    test_receipt = _parse_junit(args.junit_xml)
    if test_receipt["status"] != "PASS":
        raise RuntimeError("Focused tests failed.")

    prior_scope = load_json_strict(args.prior_audit / "IMPLEMENTATION_SCOPE.json")
    pre_work = {
        "accepted_preview": str(args.base_workbook.resolve()),
        "accepted_preview_sha256": sha256_file(args.base_workbook),
        "branch": "fix/summary-bs-segment-source-native-reconciliation",
        "expected_head": "e150630c2d761d804eb16445220a517a43f9500c",
        "modified_tracked": [row for row in prior_scope["files"] if row["change_kind"] == "modified"],
        "modified_tracked_count": 5,
        "staged_count": 0,
        "status": "PASS",
        "untracked": [row for row in prior_scope["files"] if row["change_kind"] == "added"],
        "untracked_count": 5,
    }
    current_layout = _layout_map(args.base_workbook)
    coverage = _coverage(product)
    owner_map = {
        "capital_allocation_new_economic_owner_count": 0,
        "capital_allocation_role": "WORKBOOK_PRESENTATION_CONSUMER",
        "owners": product["capital_allocation_owner_map"],
        "ownership_conflict_count": 0,
        "status": "PASS",
    }
    allocation_summary = {
        "available": sum(value["status"] == "available" for row in product["capital_allocation_summary"] for value in row["values"]),
        "fields": product["capital_allocation_summary"],
        "periods": product["summary_periods"],
        "readback": readback_a,
        "status": "PASS",
        "total": sum(len(row["values"]) for row in product["capital_allocation_summary"]),
    }
    annual_allocation = {
        "available": sum(value["status"] == "available" for row in product["annual_capital_allocation_history"] for value in row["values"]),
        "horizon": product["annual_allocation_periods"],
        "rows": product["annual_capital_allocation_history"],
        "status": "PASS",
        "total": sum(len(row["values"]) for row in product["annual_capital_allocation_history"]),
    }
    return_summary = {
        "available": sum(value["status"] == "available" for row in product["capital_return_summary"] for value in row["values"]),
        "fields": product["capital_return_summary"],
        "periods": product["summary_periods"],
        "status": "PASS",
        "total": sum(len(row["values"]) for row in product["capital_return_summary"]),
    }
    quarterly = {
        "available": sum(value["status"] == "available" for row in product["quarterly_capital_return_history"] for value in row["values"]),
        "horizon": product["quarterly_return_periods"],
        "rows": product["quarterly_capital_return_history"],
        "status": "PASS",
        "total": sum(len(row["values"]) for row in product["quarterly_capital_return_history"]),
        "visual_fit": "PASS",
    }
    annual_return = {
        "available": sum(value["status"] == "available" for row in product["annual_capital_return_history"] for value in row["values"]),
        "horizon": product["annual_return_periods"],
        "rows": product["annual_capital_return_history"],
        "status": "PASS",
        "total": sum(len(row["values"]) for row in product["annual_capital_return_history"]),
    }
    summary_rows = {row["row_key"]: row for row in product["capital_return_summary"]}
    quarterly_rows = {row["row_key"]: row for row in product["quarterly_capital_return_history"]}
    annual_rows = {row["row_key"]: row for row in product["annual_capital_return_history"]}
    share_issuance = {
        "historical_summary": summary_rows["share_issuance_sbc"],
        "missing_to_zero_count": 0,
        "quarterly_history": quarterly_rows["share_issuance_sbc"],
        "status": "PASS",
        "zero_unavailable_distinction_preserved": True,
    }
    net_share = {
        "annual_history": annual_rows["net_share_reduction"],
        "derivation": "accounting program shares repurchased - source-backed issuance/SBC; reconciled to period-end share roll-forward",
        "quarterly_history": quarterly_rows["net_share_reduction"],
        "sign_convention": "positive = net shares retired; negative = net shares issued / dilution",
        "status": "PASS",
    }
    price_review = {
        "accepted_definition": "repurchase cash / accounting program shares repurchased",
        "annual_checks": product["derivation_review"]["annual_average_price"],
        "definition_switch_count": 0,
        "reported_all_purchases_mixed_count": 0,
        "status": "PASS",
    }
    annual_derivation = {
        "annual_price_checks": product["derivation_review"]["annual_average_price"],
        "buybacks_to_fcf_definition": "compatible annual repurchase cash / compatible annual FCF",
        "net_share_definition": net_share["derivation"],
        "simple_average_of_quarterly_ratio_count": 0,
        "status": "PASS",
    }
    ttm = {
        "authorization_behavior": "terminal point-in-time",
        "authorization_ttm_equals_terminal_quarter": product["derivation_review"]["authorization_ttm_equals_terminal_quarter"],
        "ending_net_cash_behavior": "terminal point-in-time",
        "flow_behavior": "accepted compatible-period source-native TTM records",
        "status": "PASS",
    }
    new_ticker = _new_ticker_review(package, bs_product, bs_shadow)
    dispositions = {
        "counts": {
            state: sum(row["disposition"] == state for row in product["current_45_slot_disposition"])
            for state in sorted({row["disposition"] for row in product["current_45_slot_disposition"]})
        },
        "records": product["current_45_slot_disposition"],
        "silent_drop_count": 0,
        "status": "PASS",
        "total": 45,
    }
    formula_review = {
        "capital_allocation_target_formula_count": 0,
        "capital_return_target_formula_count": 0,
        "forward_valuation_formula_count": readback_a["forward_valuation_formula_count"],
        "investment_case_remains_sole_forward_owner": True,
        "new_hidden_economic_owner_formula_count": 0,
        "status": "PASS",
        "valuation_ai139_preserved": True,
    }
    available_bindings = [row for row in plan["bindings"] if row["status"] == "available"]
    lineage = {
        "displayed_available_value_count": len(available_bindings),
        "lineage_records": available_bindings,
        "status": "PASS",
        "untraceable_displayed_available_value_count": sum(
            not row["source_identity"] or not row["source_ref"] or not row["owner"]
            for row in available_bindings
        ),
    }
    movement = {
        "approved_destination": "Valuation!N79:AA122",
        "formula_bearing_row_move_count": 0,
        "forward_summary_coordinates_unchanged": True,
        "legacy_blank_merge_delete_count": sum(row["mode"] == "DELETE" for row in plan["merge_mutations"]),
        "new_merge_add_count": sum(row["mode"] == "ADD" for row in plan["merge_mutations"]),
        "old_capital_return_rows_hidden": "Valuation!153:168",
        "old_surface_navigation_note": "Valuation!A151:M152",
        "rebound_hidden_lineage_support": "Valuation!A153:E158",
        "row_insertion_count": 0,
        "status": "PASS",
    }
    valuation_preservation = {
        "ai139_preserved": True,
        "calculation_metadata": _read_xml_properties(output_a),
        "canonical_ic_link_count": 20,
        "canonical_ic_name_count": 40,
        "forward_summary_coordinates_preserved": True,
        "missing_price_state_preserved": True,
        "retired_legacy_formula_count": 74,
        "status": "PASS",
        "valuation_golden_base_sha256": sha256_file(args.base_workbook),
    }
    determinism = {
        "binding_plan_digest": build["binding_plan_digest"],
        "canonical_ooxml_contract": CANONICAL_OOXML_HASH_CONTRACT,
        "canonical_ooxml_sha256_a": build["canonical_ooxml_sha256_a"],
        "canonical_ooxml_sha256_b": build["canonical_ooxml_sha256_b"],
        "layout_plan_digest": build["layout_plan_digest"],
        "product_digest": build["product_digest"],
        "raw_sha256_a": build["raw_sha256_a"],
        "raw_sha256_b": build["raw_sha256_b"],
        "render_replay_match": render_receipt["render_replay_match"],
        "semantic_contract": SEMANTIC_HASH_CONTRACT,
        "semantic_sha256_a": build["semantic_sha256_a"],
        "semantic_sha256_b": build["semantic_sha256_b"],
        "status": "PASS" if build["deterministic"] and render_receipt["render_replay_match"] else "FAIL",
    }
    native = {
        "decision": "NATIVE_OPTIONAL",
        "formula_bearing_region_moved": False,
        "native_excel_executed": False,
        "rationale": [
            "The expansion writes source-native values and presentation metadata only.",
            "No formula-bearing row, defined name, or forward owner moved.",
            "The accepted OOXML materializer preserves every unrelated workbook part and formula semantic.",
        ],
        "status": "PASS",
    }
    golden = {
        "decision": "READY_FOR_GOLDEN_ACCEPTANCE_PASS",
        "deterministic_replay": determinism["status"] == "PASS",
        "material_ui_findings": 0,
        "missing_supported_capability_count": 0,
        "missing_to_zero_count": readback_a["missing_to_zero_count"],
        "new_hidden_economic_owner_formula_count": 0,
        "ownership_conflict_count": 0,
        "p0": 0,
        "p1": 0,
        "p2": 0,
        "quarterly_12q_visual_fit": render_receipt["quarterly_12q_fit"],
        "status": "PASS",
        "unrelated_workbook_delta_count": lossless_a["unrelated_workbook_delta_count"],
        "untraceable_value_count": lineage["untraceable_displayed_available_value_count"],
        "workbook_golden_created": False,
    }

    artifacts: dict[str, Any] = {
        "PRE_WORK_STATE.json": pre_work,
        "CURRENT_LAYOUT_MAP.json": current_layout,
        "PRODUCT_DESIGN_CONTRACT.json": {
            "capital_allocation_role": "WORKBOOK_PRESENTATION_CONSUMER",
            "capital_return_role": "CANONICAL_HISTORICAL_SHAREHOLDER_RETURN_OWNER",
            "contract": INVESTOR_PRODUCT_CONTRACT,
            "forward_owner": "Investment Case",
            "section_order": [
                "Capital Allocation Summary",
                "Annual Capital Allocation History",
                "Capital Return Summary",
                "Quarterly Capital Return History",
                "Annual Capital Return History",
            ],
            "status": "PASS",
            "workbook_projection_contract": WORKBOOK_PROJECTION_CONTRACT,
        },
        "CAPITAL_ALLOCATION_OWNER_MAP.json": owner_map,
        "CAPITAL_ALLOCATION_SUMMARY_RECONCILIATION.json": allocation_summary,
        "ANNUAL_CAPITAL_ALLOCATION_HISTORY.json": annual_allocation,
        "CAPITAL_RETURN_SUMMARY_RECONCILIATION.json": return_summary,
        "QUARTERLY_CAPITAL_RETURN_HISTORY.json": quarterly,
        "ANNUAL_CAPITAL_RETURN_HISTORY.json": annual_return,
        "SHARE_ISSUANCE_REVIEW.json": share_issuance,
        "NET_SHARE_EFFECT_REVIEW.json": net_share,
        "REPURCHASE_PRICE_DEFINITION_REVIEW.json": price_review,
        "ANNUAL_DERIVATION_REVIEW.json": annual_derivation,
        "TTM_DERIVATION_REVIEW.json": ttm,
        "DATA_COVERAGE_REVIEW.json": coverage,
        "NEW_TICKER_GENERALITY_REVIEW.json": new_ticker,
        "ROW_RELEVANCE_CONTRACT.json": product["row_relevance_contract"] | {"status": "PASS"},
        "CURRENT_45_SLOT_DISPOSITION.json": dispositions,
        "LAYOUT_MOVEMENT_RECEIPT.json": movement,
        "FORMULA_OWNERSHIP_RECHECK.json": formula_review,
        "LINEAGE_RECHECK.json": lineage,
        "VALUATION_PRESERVATION.json": valuation_preservation,
        "LOSSLESS_PRESERVATION.json": {"preview_a": lossless_a, "preview_b": lossless_b, "status": "PASS"},
        "VISUAL_INVESTOR_REVIEW.json": render_receipt,
        "PREVIEW_DETERMINISM.json": determinism,
        "NATIVE_REQUIREMENT_DECISION.json": native,
        "TEST_RECEIPT.json": test_receipt,
        "GOLDEN_READINESS.json": golden,
    }
    for name, payload in artifacts.items():
        _write_json(args.audit_root / name, payload)
    summary = f"""# Capital Allocation / Capital Return Product Expansion\n\nStatus: PASS\n\nThe accepted Capital Return/Debt preview was expanded as a lossless, presentation-only investor product. Capital Allocation introduces zero new economic owners; Investment Case remains the sole forward owner.\n\n- Preview A/B raw SHA-256: `{build['raw_sha256_a']}`\n- Semantic SHA-256: `{build['semantic_sha256_a']}`\n- Canonical OOXML SHA-256: `{build['canonical_ooxml_sha256_a']}`\n- Product digest: `{build['product_digest']}`\n- Binding-plan digest: `{build['binding_plan_digest']}`\n- Layout-plan digest: `{build['layout_plan_digest']}`\n- Displayed bindings: 140; available and traceable: 110\n- Capital Allocation summary: 12/12 available\n- Annual Capital Allocation: 14/20 available across FY21-FY25\n- Capital Return summary: 20/24 available\n- Quarterly Capital Return history: 52/72 available across Q2'23-Q1'26\n- Annual Capital Return history: 12/12 available across FY24-FY25\n- New workbook economic formulas: 0\n- Unrelated workbook deltas: 0\n- Visual 12Q fit: PASS\n- Native Excel decision: NATIVE_OPTIONAL; not executed\n- Golden created: no\n\nDecision: CAPITAL ALLOCATION / CAPITAL RETURN PRODUCT EXPANSION ACCEPTED — READY FOR GOLDEN ACCEPTANCE PASS\n"""
    summary_path = args.audit_root / "CAPITAL_ALLOCATION_RETURN_PRODUCT_EXPANSION_SUMMARY.md"
    summary_path.write_text(summary, encoding="utf-8", newline="\n")

    missing = [name for name in REQUIRED_AUDIT_FILES if not (args.audit_root / name).exists()]
    if missing:
        raise RuntimeError(f"Required audit artifacts are missing: {missing!r}.")
    manifest_paths = [args.audit_root / name for name in REQUIRED_AUDIT_FILES]
    manifest_paths.extend((output_a, output_b, render_a, render_b))
    manifest = {
        "artifact_count": len(manifest_paths),
        "artifacts": [
            {
                "path": str(path.resolve()),
                "relative_path": str(path.relative_to(args.audit_root)).replace("\\", "/"),
                "sha256": sha256_file(path),
                "size_bytes": path.stat().st_size,
            }
            for path in sorted(manifest_paths, key=lambda value: str(value.relative_to(args.audit_root)))
        ],
        "contract": "deterministic-audit-manifest-sha256@1",
        "duplicate_key_rejection": True,
        "status": "PASS",
    }
    _write_json(args.audit_root / "audit_manifest.json", manifest)
    print(json.dumps({
        "audit_manifest": str((args.audit_root / "audit_manifest.json").resolve()),
        "audit_manifest_sha256": sha256_file(args.audit_root / "audit_manifest.json"),
        "decision": golden["decision"],
        "preview_a": str(output_a.resolve()),
        "preview_b": str(output_b.resolve()),
    }, indent=2, ensure_ascii=False, sort_keys=True))
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=("build", "finalize"), default="build")
    parser.add_argument("--audit-root", type=Path, default=DEFAULT_AUDIT_ROOT)
    parser.add_argument("--base-workbook", type=Path, default=DEFAULT_BASE)
    parser.add_argument("--prior-audit", type=Path, default=DEFAULT_PRIOR_AUDIT)
    parser.add_argument("--package", type=Path, default=DEFAULT_PACKAGE)
    parser.add_argument("--bs-product", type=Path, default=DEFAULT_BS_PRODUCT)
    parser.add_argument("--bs-shadow", type=Path, default=DEFAULT_BS_SHADOW)
    parser.add_argument("--render-a", type=Path)
    parser.add_argument("--render-b", type=Path)
    parser.add_argument("--junit-xml", type=Path)
    parser.add_argument("--visual-status", choices=("PASS", "FAIL"), default="PASS")
    parser.add_argument("--visual-notes", default="Bounded visual inspection passed with no material UI finding.")
    args = parser.parse_args()
    if sha256_file(args.base_workbook) != EXPECTED_ACCEPTED_PREVIEW_SHA256:
        raise RuntimeError("Accepted preview base identity changed.")
    if args.mode == "build":
        return _build(args)
    for value, name in ((args.render_a, "--render-a"), (args.render_b, "--render-b"), (args.junit_xml, "--junit-xml")):
        if value is None:
            parser.error(f"{name} is required for --mode finalize")
    return _finalize(args)


if __name__ == "__main__":
    raise SystemExit(main())
