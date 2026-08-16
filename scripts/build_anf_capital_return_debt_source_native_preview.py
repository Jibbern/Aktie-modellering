from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
import sys
from typing import Any
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pbi_xbrl.json_schema_validation import load_json_strict
from pbi_xbrl.longitudinal_memory.capital_return_debt_workbook_projection import (
    build_capital_return_debt_workbook_projection_plan,
    materialize_capital_return_debt_workbook_projection,
)
from pbi_xbrl.longitudinal_memory.summary_bs_workbook_materialization import (
    CANONICAL_OOXML_HASH_CONTRACT,
    canonical_ooxml_sha256,
    sha256_file,
)


DEFAULT_DATA_ROOT = Path(r"C:\Users\Jibbe\Aktier\StockModelData")
DEFAULT_AUDIT_ROOT = (
    DEFAULT_DATA_ROOT / "audit" / "capital_return_debt_bounded_correction_2026-08-16"
)
DEFAULT_BASE = (
    DEFAULT_DATA_ROOT
    / "audit"
    / "valuation_golden_acceptance_2026-08-15"
    / "golden"
    / "ANF_valuation_source_native_golden_v1.xlsx"
)
DEFAULT_PACKAGE = (
    DEFAULT_DATA_ROOT
    / "outputs"
    / "stress_tests"
    / "ANF_new_ticker_engine"
    / "ANF_normalized_data_package.json"
)
SEMANTIC_HASH_CONTRACT = "capital-return-debt-pre-golden-semantic-snapshot-sha256@1"


def _canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")


def _digest(value: Any) -> str:
    return hashlib.sha256(_canonical_bytes(value)).hexdigest()


def _calculation_metadata(path: Path) -> dict[str, str]:
    namespace = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(path, "r") as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    calc = workbook.find("m:calcPr", namespace)
    if calc is None:
        raise RuntimeError("Workbook lacks calcPr metadata.")
    return dict(sorted(calc.attrib.items()))


def _semantic_snapshot(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        ranges = {
            "Valuation": ((63, 13, 13), (152, 1, 13), (153, 1, 13), (154, 1, 13),
                          (155, 1, 13), (156, 1, 13), (157, 1, 13), (158, 1, 13),
                          (159, 1, 13), (160, 1, 13), (161, 1, 13), (162, 1, 13),
                          (163, 1, 13), (164, 1, 13), (165, 1, 13), (166, 1, 13),
                          (167, 1, 13), (168, 1, 13), *(
                              (row, 30, 41) for row in range(172, 187)
                          )),
            "Debt_Profile": tuple((row, 1, 10) for row in range(1, 15)),
            "Revolver_History": tuple((row, 1, 16) for row in range(1, 16)),
            "Leverage_Liquidity": tuple((row, 1, 14) for row in range(1, 16)),
            "Debt_Credit_Notes": tuple((row, 1, 8) for row in range(1, 10)),
            "Debt_Maturity_Ladder": tuple((row, 1, 8) for row in range(1, 4)),
        }
        cells: list[dict[str, Any]] = []
        for sheet_name, row_ranges in ranges.items():
            sheet = workbook[sheet_name]
            for row, minimum_column, maximum_column in row_ranges:
                for column in range(minimum_column, maximum_column + 1):
                    cell = sheet.cell(row, column)
                    cells.append(
                        {
                            "cell": cell.coordinate,
                            "data_type": cell.data_type,
                            "number_format": cell.number_format,
                            "sheet": sheet_name,
                            "style_id": cell.style_id,
                            "value": cell.value,
                        }
                    )
        sheet_states = {
            sheet.title: sheet.sheet_state
            for sheet in workbook.worksheets
        }
        table_contracts: dict[str, dict[str, Any]] = {}
        for sheet_name in (
            "Debt_Profile",
            "Revolver_History",
            "Leverage_Liquidity",
            "Debt_Credit_Notes",
        ):
            sheet = workbook[sheet_name]
            table_contracts[sheet_name] = {}
            for name in sheet.tables.keys():
                table = sheet.tables[name]
                table_contracts[sheet_name][name] = {
                    "columns": [column.name for column in table.tableColumns],
                    "ref": table.ref,
                    "show_row_stripes": table.tableStyleInfo.showRowStripes,
                }
        return {
            "calculation_metadata": _calculation_metadata(path),
            "cells": cells,
            "contract": SEMANTIC_HASH_CONTRACT,
            "sheet_states": sheet_states,
            "table_contracts": table_contracts,
        }
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--audit-root", type=Path, default=DEFAULT_AUDIT_ROOT)
    parser.add_argument("--base-workbook", type=Path, default=DEFAULT_BASE)
    parser.add_argument("--package", type=Path, default=DEFAULT_PACKAGE)
    args = parser.parse_args()

    args.audit_root.mkdir(parents=True, exist_ok=True)
    output_a = args.audit_root / "ANF_capital_return_debt_source_native_preview_a.xlsx"
    output_b = args.audit_root / "ANF_capital_return_debt_source_native_preview_b.xlsx"
    for output in (output_a, output_b):
        if output.exists():
            raise RuntimeError(f"Refusing to overwrite existing preview: {output}.")

    package = load_json_strict(args.package)
    plan_a = build_capital_return_debt_workbook_projection_plan(
        package=package,
        source_package_path=args.package,
        base_workbook=args.base_workbook,
    )
    plan_b = build_capital_return_debt_workbook_projection_plan(
        package=load_json_strict(args.package),
        source_package_path=args.package,
        base_workbook=args.base_workbook,
    )
    if plan_a.to_dict() != plan_b.to_dict():
        raise RuntimeError("Independent binding-plan replay changed.")

    result_a = materialize_capital_return_debt_workbook_projection(
        plan=plan_a,
        base_workbook=args.base_workbook,
        output_workbook=output_a,
    )
    result_b = materialize_capital_return_debt_workbook_projection(
        plan=plan_b,
        base_workbook=args.base_workbook,
        output_workbook=output_b,
    )
    snapshot_a = _semantic_snapshot(output_a)
    snapshot_b = _semantic_snapshot(output_b)
    semantic_a = _digest(snapshot_a)
    semantic_b = _digest(snapshot_b)
    canonical_a = canonical_ooxml_sha256(output_a)
    canonical_b = canonical_ooxml_sha256(output_b)
    raw_a = sha256_file(output_a)
    raw_b = sha256_file(output_b)
    if not (
        raw_a == raw_b
        and semantic_a == semantic_b
        and canonical_a == canonical_b
        and result_a.as_dict() == result_b.as_dict()
    ):
        raise RuntimeError("Independent preview replay is not deterministic.")

    receipt = {
        "base_workbook": str(args.base_workbook.resolve()),
        "base_workbook_sha256": sha256_file(args.base_workbook),
        "binding_plan_digest": plan_a.binding_plan_digest,
        "canonical_ooxml_contract": CANONICAL_OOXML_HASH_CONTRACT,
        "canonical_ooxml_sha256": canonical_a,
        "materialization": result_a.as_dict(),
        "preview_a": str(output_a.resolve()),
        "preview_b": str(output_b.resolve()),
        "raw_sha256": raw_a,
        "semantic_contract": SEMANTIC_HASH_CONTRACT,
        "semantic_sha256": semantic_a,
        "source_package": str(args.package.resolve()),
        "source_package_sha256": sha256_file(args.package),
    }
    work = args.audit_root / "work"
    work.mkdir(parents=True, exist_ok=True)
    (work / "build_result.json").write_bytes(_canonical_bytes(receipt) + b"\n")
    print(json.dumps(receipt, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
