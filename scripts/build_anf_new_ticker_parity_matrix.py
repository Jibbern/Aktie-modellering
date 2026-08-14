"""Build the ANF legacy-oracle parity matrix for the generic new-ticker engine.

ANF is intentionally treated as a migration fixture.  The contract records
business keys, lineage, normalized paths, bindings, and generic formula
ownership; it never copies legacy workbook values into the frozen shell.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pbi_xbrl.json_schema_validation import load_json_strict
from pbi_xbrl.new_ticker_guidance_scope import guidance_scope_key, normalize_guidance_scope
from pbi_xbrl.new_ticker_investment_case_formula_surface import (
    CANONICAL_SCENARIOS,
    CANONICAL_VALUATION_METHODS,
    canonical_valuation_matrix_row,
)
from pbi_xbrl.standard_template_formula_contract import (
    FORMULA_CONTRACT_VERSION,
    FORMULA_ROWS,
)
from pbi_xbrl.standard_template_audit_freshness import _portable_file_sha256
from scripts.build_anf_shadow_normalized_package import (
    _anf_history_source_evidence,
    _anf_legacy_guidance_horizon,
)


DEFAULT_SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
DEFAULT_BINDINGS = ROOT / "docs" / "workbook_binding_map.json"
DEFAULT_OUTPUT_JSON = ROOT / "docs" / "anf_new_ticker_parity_matrix.json"
DEFAULT_OUTPUT_MD = ROOT / "docs" / "anf_new_ticker_parity_matrix.md"

GUIDANCE_MODULE_ID = "guidance_promises"
DERIVED_GUIDANCE_ROOT = "_derived_workbook.guidance."
GUIDANCE_SOURCE_COLLECTION = "normalized_guidance.items"
GUIDANCE_PROJECTION_ID = "valuation_guidance_projection"
GUIDANCE_PROJECTION_AUTHORITY = (
    "pbi_xbrl.new_ticker_guidance_scope.build_valuation_guidance_projection"
)

LEGACY_REPORT_PLACEHOLDER_ROWS = {
    "total_debt": ("REPORT_BS_Q", "Total debt"),
    "debt_core": ("REPORT_BS_Q", "Debt core"),
    "interest_paid": ("REPORT_CF_Q", "Cash interest"),
}

QUARTERLY_CORE = {
    "revenue",
    "gross_profit",
    "operating_income",
    "base_ebitda",
    "adjusted_ebitda",
    "net_income",
    "operating_cash_flow",
    "capital_expenditures",
    "diluted_shares",
    "eps",
    "adjusted_eps",
}
QUARTERLY_BALANCE_SHEET = {
    "cash",
    "marketable_securities",
    "accounts_receivable",
    "inventory",
    "current_assets",
    "property_plant_equipment_net",
    "goodwill",
    "intangibles",
    "total_assets",
    "accounts_payable",
    "accrued_liabilities",
    "debt_current",
    "lease_liabilities_current",
    "current_liabilities",
    "debt_core",
    "lease_liabilities_noncurrent",
    "pension_obligation_net",
    "other_liabilities_noncurrent",
    "total_liabilities",
    "total_equity",
    "shares_outstanding",
}
ANNUAL_CORE = {
    "revenue",
    "gross_profit",
    "operating_income",
    "base_ebitda",
    "adjusted_ebitda",
    "net_income",
    "operating_cash_flow",
    "capital_expenditures",
    "diluted_shares",
    "total_equity",
    "cash",
    "debt_core",
}

# This map is the independent legacy inventory contract.  The keys are actual
# History_Q headers, not normalized-package field names.
LEGACY_HISTORY_METRICS: dict[str, tuple[str, str, str]] = {
    "revenue": ("revenue", "$m", "quarterly_financials"),
    "cogs": ("cost_of_goods_sold", "$m", "quarterly_financials"),
    "gross_profit": ("gross_profit", "$m", "quarterly_financials"),
    "op_income": ("operating_income", "$m", "quarterly_financials"),
    "net_income": ("net_income", "$m", "quarterly_financials"),
    "cfo": ("operating_cash_flow", "$m", "cash_flow"),
    "capex": ("capital_expenditures", "$m", "cash_flow"),
    "tax_paid": ("income_taxes_paid", "$m", "cash_flow"),
    "da": ("depreciation_amortization", "$m", "cash_flow"),
    "interest_paid": ("interest_paid", "$m", "cash_flow"),
    "interest_expense_net": ("interest_expense", "$m", "quarterly_financials"),
    "buybacks_cash": ("buybacks_cash", "$m", "cash_flow"),
    "dividends_cash": ("dividends_cash", "$m", "cash_flow"),
    "acquisitions_cash": ("acquisitions_cash", "$m", "cash_flow"),
    "debt_repayment": ("debt_repayment", "$m", "cash_flow"),
    "debt_issuance": ("debt_issuance", "$m", "cash_flow"),
    "cash": ("cash", "$m", "balance_sheet"),
    "short_term_investments": ("short_term_investments", "$m", "balance_sheet"),
    "assets": ("total_assets", "$m", "balance_sheet"),
    "liabilities": ("total_liabilities", "$m", "balance_sheet"),
    "assets_current": ("current_assets", "$m", "balance_sheet"),
    "liabilities_current": ("current_liabilities", "$m", "balance_sheet"),
    "accounts_receivable": ("accounts_receivable", "$m", "balance_sheet"),
    "inventory": ("inventory", "$m", "balance_sheet"),
    "accounts_payable_current": ("accounts_payable", "$m", "balance_sheet"),
    "accrued_liabilities_current": ("accrued_liabilities", "$m", "balance_sheet"),
    "debt_current": ("debt_current", "$m", "balance_sheet"),
    "property_plant_equipment_net": ("property_plant_equipment_net", "$m", "balance_sheet"),
    "other_assets_noncurrent": ("other_assets_noncurrent", "$m", "balance_sheet"),
    "other_liabilities_noncurrent": ("other_liabilities_noncurrent", "$m", "balance_sheet"),
    "total_equity": ("total_equity", "$m", "balance_sheet"),
    "goodwill": ("goodwill", "$m", "balance_sheet"),
    "intangibles": ("intangibles", "$m", "balance_sheet"),
    "shares_outstanding": ("shares_outstanding", "m shares", "per_share"),
    "pension_obligation_net": ("pension_obligation_net", "$m", "balance_sheet"),
    "total_debt": ("total_debt", "$m", "balance_sheet"),
    "debt_core": ("debt_core", "$m", "balance_sheet"),
    "lease_liabilities": ("lease_liabilities", "$m", "balance_sheet"),
    "shares_diluted": ("diluted_shares", "m shares", "per_share"),
    "ebitda": ("base_ebitda", "$m", "quarterly_financials"),
    "eps_diluted": ("eps", "$/share", "per_share"),
    "lease_liabilities_current": ("lease_liabilities_current", "$m", "balance_sheet"),
    "lease_liabilities_noncurrent": ("lease_liabilities_noncurrent", "$m", "balance_sheet"),
    "marketable_securities": ("marketable_securities", "$m", "balance_sheet"),
    "operating_margin": ("operating_margin", "%", "formula_derived_metrics"),
}

LEGACY_DERIVED_HISTORY_METRICS = {"operating_margin"}

ANNUAL_FLOW_FIELDS = {
    "revenue", "cost_of_goods_sold", "gross_profit", "operating_income", "base_ebitda", "net_income",
    "operating_cash_flow", "capital_expenditures", "interest_paid", "interest_expense",
    "income_taxes_paid", "depreciation_amortization", "buybacks_cash", "dividends_cash",
    "acquisitions_cash", "debt_repayment", "debt_issuance",
}
ANNUAL_POINT_IN_TIME_FIELDS = {
    "cash", "short_term_investments", "total_assets", "total_liabilities", "current_assets",
    "current_liabilities", "accounts_receivable", "inventory", "accounts_payable",
    "accrued_liabilities", "debt_current", "property_plant_equipment_net", "other_assets_noncurrent",
    "other_liabilities_noncurrent", "total_equity", "goodwill", "intangibles",
    "shares_outstanding", "pension_obligation_net", "total_debt", "debt_core",
    "lease_liabilities", "lease_liabilities_current", "lease_liabilities_noncurrent",
    "marketable_securities",
}

SCALAR_REQUIREMENTS = (
    ("valuation_inputs", "valuation_inputs.price", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.as_of_date", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.shares_outstanding", "unavailable_missing_evidence"),
    ("valuation_inputs", "valuation_inputs.diluted_shares", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.net_debt", "unavailable_missing_evidence"),
    ("valuation_inputs", "valuation_inputs.base_ebitda_ttm", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.adjusted_ebitda_ttm", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.revenue_ttm", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.operating_cash_flow_ttm", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.free_cash_flow_ttm", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.capex_ttm", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.eps_ttm", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.adjusted_eps_ttm", "unavailable_missing_evidence"),
    ("valuation_inputs", "valuation_inputs.book_value_per_share", "unavailable_missing_evidence"),
    ("valuation_inputs", "valuation_inputs.tangible_book_value_per_share", "unavailable_missing_evidence"),
    ("valuation_inputs", "valuation_inputs.interest_paid_ttm", "unavailable_missing_evidence"),
    ("valuation_inputs", "valuation_inputs.adjusted_fcf_ttm", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.target_ev_adjusted_ebitda", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.target_ev_ebitda", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.target_ev_yield", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.maintenance_capex_ratio", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.recurring_cash_costs", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.working_capital_normalization", "intentionally_rejected"),
    ("valuation_inputs", "valuation_inputs.per_share_denominator", "intentionally_rejected"),
)

RETIRED_DUPLICATE_VALUATION_INPUT_PATHS = frozenset(
    {
        "valuation_inputs.as_of_date",
        "valuation_inputs.diluted_shares",
        "valuation_inputs.base_ebitda_ttm",
        "valuation_inputs.adjusted_ebitda_ttm",
        "valuation_inputs.revenue_ttm",
        "valuation_inputs.operating_cash_flow_ttm",
        "valuation_inputs.free_cash_flow_ttm",
        "valuation_inputs.capex_ttm",
        "valuation_inputs.eps_ttm",
    }
)

# These keys are inventoried from stable visible legacy business rows before the
# normalized package is consulted. Text may be improved when the legacy wording
# contains unsupported claims or implementation language.
LEGACY_NARRATIVE_SCALARS = (
    ("summary:company-description", "summary", "SUMMARY", "A3", "company_profile.business_description", "may_improve_semantically"),
    ("summary:strategic-context", "summary", "SUMMARY", "A5", "company_profile.strategic_context", "may_improve_semantically"),
    ("summary:key-advantage", "summary", "SUMMARY", "A7", "company_profile.key_advantages", "may_improve_semantically"),
    ("investment:title", "investment_case", "ANF_Investment_Case", "A1", "ticker_metadata.investment_case_title", "must_reproduce"),
    ("investment:model-read", "investment_case", "ANF_Investment_Case", "B5", "investment_case.summary", "may_improve_semantically"),
    ("investment:why-it-can-work", "investment_case", "ANF_Investment_Case", "B6", "investment_case.why_it_can_work", "may_improve_semantically"),
    ("investment:key-debate", "investment_case", "ANF_Investment_Case", "B7", "investment_case.key_debate", "may_improve_semantically"),
    ("investment:upside", "investment_case", "ANF_Investment_Case", "B8", "investment_case.upside_factors", "may_improve_semantically"),
    ("investment:downside", "investment_case", "ANF_Investment_Case", "B9", "investment_case.downside_factors", "may_improve_semantically"),
    ("investment:watch-next", "investment_case", "ANF_Investment_Case", "B10", "investment_case.watch_next", "may_improve_semantically"),
    ("investment:current-stance", "investment_case", "ANF_Investment_Case", "B11", "investment_case.current_stance", "may_improve_semantically"),
    ("drivers:current-actual-read", "operating_drivers", "Operating_Drivers", "B13", "operating_drivers.current_outlook.current_actual_read", "may_improve_semantically"),
    ("drivers:current-actual-use", "operating_drivers", "Operating_Drivers", "H13", "operating_drivers.current_outlook.current_actual_use", "may_improve_semantically"),
    ("drivers:current-guidance-read", "operating_drivers", "Operating_Drivers", "B14", "operating_drivers.current_outlook.current_guidance_read", "may_improve_semantically"),
    ("drivers:current-guidance-use", "operating_drivers", "Operating_Drivers", "H14", "operating_drivers.current_outlook.current_guidance_use", "may_improve_semantically"),
    ("drivers:margin-bridge-read", "operating_drivers", "Operating_Drivers", "B15", "operating_drivers.current_outlook.margin_bridge_read", "may_improve_semantically"),
    ("drivers:margin-bridge-use", "operating_drivers", "Operating_Drivers", "H15", "operating_drivers.current_outlook.margin_bridge_use", "may_improve_semantically"),
    ("quarter-notes:model-read", "quarter_notes", "Quarter_Notes_UI", "B3", "quarter_notes.summary.model_read", "may_improve_semantically"),
    ("quarter-notes:what-changed", "quarter_notes", "Quarter_Notes_UI", "B4", "quarter_notes.summary.what_changed", "may_improve_semantically"),
    ("quarter-notes:watch-next", "quarter_notes", "Quarter_Notes_UI", "B5", "quarter_notes.summary.watch_next", "may_improve_semantically"),
    ("quarter-notes:key-caveat", "quarter_notes", "Quarter_Notes_UI", "B6", "quarter_notes.summary.key_caveat", "may_improve_semantically"),
)


def _default_data_root() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = parent / "StockModelData"
        if candidate.exists():
            return candidate
    return ROOT.parent / "StockModelData"


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _path_get(payload: Mapping[str, Any], path: str) -> Any:
    current: Any = payload
    for part in path.split("."):
        if not isinstance(current, Mapping) or part not in current:
            return None
        current = current[part]
    return current


def _field_state(field: Any) -> tuple[Any, str, str]:
    if not isinstance(field, Mapping):
        return None, "missing_source", ""
    return field.get("value"), str(field.get("status") or "missing_source"), str(field.get("source_ref") or "")


def _scalar(value: Any) -> Any:
    return value.get("value") if isinstance(value, Mapping) else value


def _write_index(plan: Mapping[str, Any]) -> dict[str, list[Mapping[str, Any]]]:
    result: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for write in plan.get("planned_writes") or []:
        if isinstance(write, Mapping):
            result[str(write.get("normalized_path") or "")].append(write)
    return result


def _planned_write_sort_key(write: Mapping[str, Any]) -> tuple[Any, ...]:
    target_cell = str(write.get("target_cell") or "")
    match = re.fullmatch(r"([A-Z]+)(\d+)", target_cell)
    column = match.group(1) if match else target_cell
    row = int(match.group(2)) if match else 0
    return (
        str(write.get("target_sheet") or ""),
        row,
        column,
        str(write.get("binding_id") or ""),
        str(write.get("target_role") or ""),
        str(write.get("normalized_path") or ""),
    )


def _active_guidance_bindings(
    binding_document: Mapping[str, Any],
) -> dict[str, Mapping[str, Any]]:
    result: dict[str, Mapping[str, Any]] = {}
    for binding in binding_document.get("bindings") or []:
        if not isinstance(binding, Mapping):
            continue
        if str(binding.get("module_id") or "") != GUIDANCE_MODULE_ID:
            continue
        if str(binding.get("planning_state") or "") != "active":
            continue
        binding_id = str(binding.get("binding_id") or "")
        if not binding_id:
            raise ValueError("Active Guidance binding is missing binding_id.")
        if binding_id in result:
            raise ValueError(f"Duplicate active Guidance binding_id: {binding_id}")
        result[binding_id] = binding
    return result


def _guidance_projection_plan(plan: Mapping[str, Any]) -> Mapping[str, Any]:
    matches = [
        row
        for row in plan.get("derived_plans") or []
        if isinstance(row, Mapping)
        and str(row.get("plan_id") or "") == GUIDANCE_PROJECTION_ID
    ]
    if len(matches) != 1:
        raise ValueError(
            "Derived Guidance bindings require exactly one "
            f"{GUIDANCE_PROJECTION_ID!r} plan; found {len(matches)}."
        )
    return matches[0]


def _guidance_destination_lineage(
    plan: Mapping[str, Any],
    binding_document: Mapping[str, Any],
) -> dict[str, Any]:
    active_bindings = _active_guidance_bindings(binding_document)
    binding_index = {
        str(binding.get("binding_id") or ""): binding
        for binding in binding_document.get("bindings") or []
        if isinstance(binding, Mapping) and binding.get("binding_id")
    }
    plan_writes = [
        write
        for write in plan.get("planned_writes") or []
        if isinstance(write, Mapping)
    ]
    for write in plan_writes:
        binding = binding_index.get(str(write.get("binding_id") or ""))
        if (
            binding is not None
            and str(binding.get("module_id") or "") == GUIDANCE_MODULE_ID
            and str(binding.get("planning_state") or "") != "active"
        ):
            raise ValueError(
                "Inactive Guidance binding produced a planned write: "
                f"{write.get('binding_id')}"
            )

    derived_bindings = {
        binding_id
        for binding_id, binding in active_bindings.items()
        if str((binding.get("row_selector") or {}).get("source_path") or "").startswith(
            DERIVED_GUIDANCE_ROOT
        )
    }
    projection_plan = _guidance_projection_plan(plan) if derived_bindings else {}
    profile_id = str(binding_document.get("module_profile_id") or "")
    seen_destinations: dict[str, str] = {}
    lineage_bindings: list[dict[str, Any]] = []

    for binding_id, binding in sorted(active_bindings.items()):
        row_selector = binding.get("row_selector") or {}
        source_path = str(row_selector.get("source_path") or binding.get("normalized_field") or "")
        derived = source_path.startswith(DERIVED_GUIDANCE_ROOT)
        binding_writes = sorted(
            (
                write
                for write in plan_writes
                if str(write.get("binding_id") or "") == binding_id
            ),
            key=_planned_write_sort_key,
        )
        destination_rows: list[dict[str, Any]] = []
        for write in binding_writes:
            destination = f"{write.get('target_sheet')}!{write.get('target_cell')}"
            previous_owner = seen_destinations.get(destination)
            if previous_owner is not None:
                raise ValueError(
                    "Duplicate active Guidance destination "
                    f"{destination}: {previous_owner} and {binding_id}"
                )
            seen_destinations[destination] = binding_id
            destination_rows.append(
                {
                    "destination": destination,
                    "normalized_path": str(write.get("normalized_path") or ""),
                    "row_key": str(write.get("row_key") or ""),
                    "source_ref": str(write.get("source_ref") or ""),
                    "target_role": str(write.get("target_role") or ""),
                }
            )

        lineage_bindings.append(
            {
                "binding_id": binding_id,
                "module_id": GUIDANCE_MODULE_ID,
                "module_profile_id": profile_id,
                "destination_sheet": str(binding.get("sheet") or ""),
                "declared_target": str(binding.get("planner_target") or binding.get("target") or ""),
                "source_selector_type": (
                    "derived_resolved_rowset" if derived else "direct_package_collection"
                ),
                "source_selector": source_path,
                "normalized_collection_root": (
                    GUIDANCE_SOURCE_COLLECTION if derived else source_path
                ),
                "resolved_rowset_producer": (
                    str(projection_plan.get("plan_id") or "") if derived else ""
                ),
                "resolver_projection_authority": (
                    GUIDANCE_PROJECTION_AUTHORITY if derived else ""
                ),
                "formula_or_value_ownership": "value_binding",
                "planning_state": str(binding.get("planning_state") or ""),
                "projected_row_count": len(
                    {
                        str(write.get("row_key") or "")
                        for write in binding_writes
                        if write.get("row_key")
                    }
                ),
                "product_role": str(binding.get("rowset_id") or binding.get("section") or ""),
                "destination_count": len(destination_rows),
                "destinations": destination_rows,
            }
        )

    return {
        "module_id": GUIDANCE_MODULE_ID,
        "module_profile_id": profile_id,
        "active_binding_count": len(lineage_bindings),
        "destination_count": len(seen_destinations),
        "bindings": lineage_bindings,
    }


def _derived_guidance_value_write_index(
    plan: Mapping[str, Any],
    binding_document: Mapping[str, Any],
) -> dict[str, list[Mapping[str, Any]]]:
    active_bindings = _active_guidance_bindings(binding_document)
    derived_binding_ids = {
        binding_id
        for binding_id, binding in active_bindings.items()
        if str((binding.get("row_selector") or {}).get("source_path") or "").startswith(
            DERIVED_GUIDANCE_ROOT
        )
    }
    result: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for write in plan.get("planned_writes") or []:
        if not isinstance(write, Mapping):
            continue
        if str(write.get("binding_id") or "") not in derived_binding_ids:
            continue
        if not str(write.get("normalized_path") or "").endswith(".value"):
            continue
        source_ref = str(write.get("source_ref") or "")
        if source_ref:
            result[source_ref].append(write)
    return {
        source_ref: sorted(writes, key=_planned_write_sort_key)
        for source_ref, writes in result.items()
    }


def _guidance_item_source_refs(item: Mapping[str, Any]) -> set[str]:
    refs = {
        str(item.get("source_ref") or ""),
        *(str(ref or "") for ref in item.get("evidence_refs") or []),
    }
    return {ref for ref in refs if ref}


def _merge_guidance_value_writes(
    direct_writes: Sequence[Mapping[str, Any]],
    item: Mapping[str, Any],
    derived_writes_by_source_ref: Mapping[str, Sequence[Mapping[str, Any]]],
) -> list[Mapping[str, Any]]:
    writes = [
        *direct_writes,
        *(
            write
            for source_ref in sorted(_guidance_item_source_refs(item))
            for write in derived_writes_by_source_ref.get(source_ref, [])
        ),
    ]
    unique: dict[tuple[str, str, str], Mapping[str, Any]] = {}
    for write in writes:
        key = (
            str(write.get("binding_id") or ""),
            str(write.get("target_sheet") or ""),
            str(write.get("target_cell") or ""),
        )
        unique[key] = write
    return sorted(unique.values(), key=_planned_write_sort_key)


def _destinations(writes: Sequence[Mapping[str, Any]]) -> list[str]:
    return sorted({f"{row.get('target_sheet')}!{row.get('target_cell')}" for row in writes})


def _binding_ids(writes: Sequence[Mapping[str, Any]]) -> list[str]:
    return sorted({str(row.get("binding_id") or "") for row in writes if row.get("binding_id")})


def _normalized_legacy_value(value: Any, unit: str) -> float | None:
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        return None
    result = float(value)
    if unit in {"$m", "m shares"}:
        result /= 1_000_000
    return round(result, 6)


def _values_match(left: Any, right: Any) -> bool:
    if isinstance(left, (int, float)) and not isinstance(left, bool) and isinstance(right, (int, float)) and not isinstance(right, bool):
        return abs(float(left) - float(right)) <= 1e-5
    return left == right


def _financial_dispositions(binding_document: Mapping[str, Any]) -> dict[tuple[str, str], Mapping[str, Any]]:
    return {
        (str(row.get("section") or ""), str(row.get("field") or "")): row
        for row in binding_document.get("financial_field_dispositions") or []
        if isinstance(row, Mapping)
    }


def _effective_disposition(
    *,
    static_disposition: Mapping[str, Any] | None,
    selector_exclusion: Mapping[str, Any] | None,
    writes: Sequence[Mapping[str, Any]],
) -> tuple[Mapping[str, Any] | None, str]:
    if static_disposition is not None:
        return static_disposition, str(static_disposition.get("disposition") or "")
    if selector_exclusion is not None:
        return selector_exclusion, str(selector_exclusion.get("disposition") or "explicitly_excluded")
    if writes:
        return None, "planned_binding"
    return None, ""


def _legacy_history_rows(legacy_path: Path) -> tuple[list[dict[str, Any]], dict[str, int]]:
    wb = load_workbook(legacy_path, read_only=True, data_only=True)
    try:
        ws = wb["History_Q"]
        headers = {str(ws.cell(1, column).value or ""): column for column in range(1, ws.max_column + 1)}
        rows: list[dict[str, Any]] = []
        for row_number in range(2, ws.max_row + 1):
            period = str(ws.cell(row_number, headers.get("fiscal_label", 0)).value or "") if headers.get("fiscal_label") else ""
            fiscal_year = ws.cell(row_number, headers.get("fiscal_year", 0)).value if headers.get("fiscal_year") else None
            fiscal_quarter = ws.cell(row_number, headers.get("fiscal_quarter", 0)).value if headers.get("fiscal_quarter") else None
            if not period or not isinstance(fiscal_year, (int, float)) or not isinstance(fiscal_quarter, (int, float)):
                continue
            rows.append(
                {
                    "row_number": row_number,
                    "period": period,
                    "fiscal_year": int(fiscal_year),
                    "fiscal_quarter": int(fiscal_quarter),
                    "values": {header: ws.cell(row_number, column).value for header, column in headers.items()},
                }
            )
        rows.sort(key=lambda row: (int(row["fiscal_year"]), int(row["fiscal_quarter"])))
        return rows, headers
    finally:
        wb.close()


def _legacy_date_key(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "")[:10]


def _legacy_unsupported_zero_placeholders(
    legacy_path: Path,
    history_rows: Sequence[Mapping[str, Any]],
) -> dict[tuple[str, str], dict[str, Any]]:
    period_by_end = {
        _legacy_date_key(row.get("values", {}).get("quarter")): str(row.get("period") or "")
        for row in history_rows
        if isinstance(row.get("values"), Mapping)
    }
    wb = load_workbook(legacy_path, read_only=True, data_only=True)
    try:
        result: dict[tuple[str, str], dict[str, Any]] = {}
        for metric, (sheet_name, line_item) in LEGACY_REPORT_PLACEHOLDER_ROWS.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            row_number = next(
                (
                    row
                    for row in range(4, ws.max_row + 1)
                    if str(ws.cell(row, 2).value or "").strip().casefold() == line_item.casefold()
                ),
                None,
            )
            if row_number is None:
                continue
            source_status = str(ws.cell(row_number, 3).value or "").strip()
            qa_status = str(ws.cell(row_number, 4).value or "").strip()
            if source_status.casefold() != "missing" and qa_status.casefold() != "fail":
                continue
            for column in range(7, ws.max_column + 1):
                value = ws.cell(row_number, column).value
                if (
                    not isinstance(value, (int, float))
                    or isinstance(value, bool)
                    or float(value) != 0.0
                ):
                    continue
                period_end = _legacy_date_key(ws.cell(3, column).value)
                period = period_by_end.get(period_end, "")
                if not period:
                    continue
                result[(metric, period)] = {
                    "metric": metric,
                    "period": period,
                    "period_end": period_end,
                    "line_item": line_item,
                    "source_status": source_status,
                    "qa_status": qa_status,
                    "value_source_ref": (
                        f"{legacy_path.name}!{sheet_name}!{get_column_letter(column)}{row_number}"
                    ),
                    "metadata_source_ref": f"{legacy_path.name}!{sheet_name}!C{row_number}:D{row_number}",
                }
        return result
    finally:
        wb.close()


def _package_row_index(package: Mapping[str, Any], section: str) -> dict[str, tuple[int, Mapping[str, Any]]]:
    return {
        str(row.get("period") or ""): (index, row)
        for index, row in enumerate(_path_get(package, f"{section}.rows") or [])
        if isinstance(row, Mapping)
    }


def _legacy_valuation_series(legacy_path: Path, row_number: int) -> dict[str, tuple[float, str]]:
    wb = load_workbook(legacy_path, read_only=True, data_only=True)
    try:
        ws = wb["Valuation"]
        result: dict[str, tuple[float, str]] = {}
        for column in range(2, 14):
            period = str(ws.cell(6, column).value or "")
            value = ws.cell(row_number, column).value
            if period and isinstance(value, (int, float)) and not isinstance(value, bool):
                result[period] = (float(value), f"{legacy_path.name}!Valuation!{get_column_letter(column)}{row_number}")
        return result
    finally:
        wb.close()


def _selector_exclusion_index(plan: Mapping[str, Any]) -> dict[str, Mapping[str, Any]]:
    result: dict[str, Mapping[str, Any]] = {}
    for binding in plan.get("bindings") or []:
        if not isinstance(binding, Mapping):
            continue
        for skipped in binding.get("skipped_rows") or []:
            if not isinstance(skipped, Mapping):
                continue
            reason = str(skipped.get("reason") or "")
            path = str(skipped.get("normalized_path") or "")
            if path and (
                reason.startswith("row_selector_")
                or reason == "period_axis_outside_visible_window"
                or reason.startswith("selector_exclusion_audit_only:")
            ):
                result.setdefault(path, skipped)
    return result


def _history_evidence_index(package: Mapping[str, Any]) -> tuple[dict[str, int], dict[tuple[str, str, int], float]]:
    periods: dict[str, int] = {}
    values: dict[tuple[str, str, int], float] = {}
    for item in _path_get(package, "calculation_history.quarterly_items") or []:
        if not isinstance(item, Mapping) or str(item.get("status") or "") != "populated":
            continue
        period = str(item.get("period") or "")
        metric = str(item.get("metric") or "")
        unit = str(item.get("unit") or "")
        ordinal = item.get("period_ordinal")
        value = item.get("value")
        source_ref = str(item.get("source_ref") or "")
        if not period or not metric or not unit or not source_ref:
            continue
        if not isinstance(ordinal, int) or not isinstance(value, (int, float)) or isinstance(value, bool):
            continue
        periods[period] = ordinal
        values[(metric, unit, ordinal)] = float(value)
    return periods, values


def _quarter_formula_calculability(
    formula_id: str,
    period: str,
    *,
    period_ordinals: Mapping[str, int],
    history_values: Mapping[tuple[str, str, int], float],
) -> tuple[bool, str]:
    ordinal = period_ordinals.get(period)
    if ordinal is None:
        return False, f"No source-backed calculation-history ordinal exists for {period}."

    def point(metric: str, unit: str, offset: int = 0, *, nonzero: bool = False) -> bool:
        value = history_values.get((metric, unit, ordinal + offset))
        return value is not None and (not nonzero or value != 0)

    def window(metric: str, unit: str, offset: int = 0, *, nonzero_total: bool = False) -> bool:
        values = [history_values.get((metric, unit, ordinal + offset - step)) for step in range(4)]
        return all(value is not None for value in values) and (not nonzero_total or sum(float(value) for value in values if value is not None) != 0)

    single_ttm = {
        "revenue_ttm": ("revenue", "$m"),
        "ebitda_ttm": ("base_ebitda", "$m"),
        "adjusted_ebitda_ttm": ("adjusted_ebitda", "$m"),
        "operating_income_ttm": ("operating_income", "$m"),
        "net_income_ttm": ("net_income", "$m"),
        "buybacks_ttm": ("buybacks_cash", "$m"),
        "dividends_ttm": ("dividends_cash", "$m"),
        "acquisitions_ttm": ("acquisitions_cash", "$m"),
        "debt_repayment_ttm": ("debt_repayment", "$m"),
        "debt_issuance_ttm": ("debt_issuance", "$m"),
        "gaap_eps_ttm": ("eps", "$/share"),
        "adjusted_eps_ttm": ("adjusted_eps", "$/share"),
        "ebitda_ttm_copy": ("base_ebitda", "$m"),
        "adjusted_ebitda_ttm_copy": ("adjusted_ebitda", "$m"),
        "operating_cash_flow_ttm": ("operating_cash_flow", "$m"),
    }
    if formula_id in single_ttm:
        metric, unit = single_ttm[formula_id]
        ok = window(metric, unit)
        return ok, "Four consecutive compatible source-backed quarters are available." if ok else f"Four consecutive {metric} quarters are unavailable."

    single_yoy = {
        "revenue_yoy": ("revenue", "$m"),
        "ebitda_yoy": ("base_ebitda", "$m"),
        "adjusted_ebitda_yoy": ("adjusted_ebitda", "$m"),
        "net_income_yoy": ("net_income", "$m"),
        "gaap_eps_yoy": ("eps", "$/share"),
    }
    if formula_id in single_yoy:
        metric, unit = single_yoy[formula_id]
        ok = point(metric, unit) and point(metric, unit, -4, nonzero=True)
        return ok, "Current and prior-year source-backed points are available." if ok else f"Current or prior-year {metric} is unavailable or has an invalid denominator."

    point_ratios = {
        "gross_margin": ("gross_profit", "revenue"),
        "operating_margin": ("operating_income", "revenue"),
        "ebitda_margin": ("base_ebitda", "revenue"),
        "adjusted_ebitda_margin": ("adjusted_ebitda", "revenue"),
        "operating_income_margin": ("operating_income", "revenue"),
        "net_margin": ("net_income", "revenue"),
        "capex_margin": ("capital_expenditures", "revenue"),
    }
    if formula_id in point_ratios:
        numerator, denominator = point_ratios[formula_id]
        ok = point(numerator, "$m") and point(denominator, "$m", nonzero=True)
        return ok, "Compatible current-period numerator and denominator are available." if ok else f"Current {numerator} or non-zero {denominator} is unavailable."

    ttm_ratios = {
        "operating_margin_ttm": ("operating_income", "revenue"),
        "ebitda_margin_ttm": ("base_ebitda", "revenue"),
        "adjusted_ebitda_margin_ttm": ("adjusted_ebitda", "revenue"),
        "operating_income_margin_ttm": ("operating_income", "revenue"),
        "net_margin_ttm": ("net_income", "revenue"),
        "capex_margin_ttm": ("capital_expenditures", "revenue"),
        "interest_coverage": ("operating_income", "interest_expense"),
        "cash_interest_coverage": ("base_ebitda", "interest_paid"),
    }
    if formula_id in ttm_ratios:
        numerator, denominator = ttm_ratios[formula_id]
        ok = window(numerator, "$m") and window(denominator, "$m", nonzero_total=True)
        return ok, "Both trailing-four-quarter series are complete and the denominator is non-zero." if ok else f"Complete compatible TTM {numerator}/{denominator} evidence is unavailable."

    if formula_id == "adjusted_ebitda_delta":
        ok = point("adjusted_ebitda", "$m") and point("base_ebitda", "$m")
    elif formula_id == "free_cash_flow":
        ok = point("operating_cash_flow", "$m") and point("capital_expenditures", "$m")
    elif formula_id == "free_cash_flow_yoy_delta":
        ok = all(point(metric, "$m", offset) for metric in ("operating_cash_flow", "capital_expenditures") for offset in (0, -4))
    elif formula_id == "free_cash_flow_ttm":
        ok = window("operating_cash_flow", "$m") and window("capital_expenditures", "$m")
    elif formula_id == "free_cash_flow_ttm_yoy_delta":
        ok = all(window(metric, "$m", offset) for metric in ("operating_cash_flow", "capital_expenditures") for offset in (0, -4))
    elif formula_id == "free_cash_flow_margin":
        ok = point("operating_cash_flow", "$m") and point("capital_expenditures", "$m") and point("revenue", "$m", nonzero=True)
    elif formula_id == "free_cash_flow_margin_ttm":
        ok = window("operating_cash_flow", "$m") and window("capital_expenditures", "$m") and window("revenue", "$m", nonzero_total=True)
    elif formula_id in {"net_debt", "core_net_cash"}:
        ok = point("debt_core", "$m") and point("cash", "$m")
    elif formula_id in {"net_debt_qoq", "diluted_shares_qoq"}:
        metric_set = (("debt_core", "$m"), ("cash", "$m")) if formula_id == "net_debt_qoq" else (("diluted_shares", "m shares"),)
        ok = all(point(metric, unit, offset) for metric, unit in metric_set for offset in (0, -1))
    elif formula_id in {"net_debt_yoy", "diluted_shares_yoy"}:
        metric_set = (("debt_core", "$m"), ("cash", "$m")) if formula_id == "net_debt_yoy" else (("diluted_shares", "m shares"),)
        ok = all(point(metric, unit, offset) for metric, unit in metric_set for offset in (0, -4))
    elif formula_id == "net_cash_with_securities":
        ok = all(point(metric, "$m") for metric in ("cash", "marketable_securities", "debt_core"))
    elif formula_id == "lease_adjusted_net_debt":
        ok = all(point(metric, "$m") for metric in ("cash", "debt_core", "lease_liabilities"))
    elif formula_id == "lease_adjusted_net_debt_with_securities":
        ok = all(point(metric, "$m") for metric in ("cash", "marketable_securities", "debt_core", "lease_liabilities"))
    elif formula_id == "net_leverage":
        ok = point("debt_core", "$m") and point("cash", "$m") and window("base_ebitda", "$m", nonzero_total=True)
    elif formula_id == "adjusted_net_leverage":
        ok = point("debt_core", "$m") and point("cash", "$m") and window("adjusted_ebitda", "$m", nonzero_total=True)
    elif formula_id == "fcf_conversion":
        ok = window("operating_cash_flow", "$m") and window("capital_expenditures", "$m") and window("base_ebitda", "$m", nonzero_total=True)
    elif formula_id == "book_value_per_share":
        ok = point("total_equity", "$m") and point("shares_outstanding", "m shares", nonzero=True)
    elif formula_id == "tangible_book_value_per_share":
        ok = all(point(metric, "$m") for metric in ("total_equity", "goodwill", "intangibles")) and point("shares_outstanding", "m shares", nonzero=True)
    elif formula_id == "free_cash_flow_per_share":
        ok = window("operating_cash_flow", "$m") and window("capital_expenditures", "$m") and point("diluted_shares", "m shares", nonzero=True)
    else:
        return False, f"No economic dependency contract is defined for formula {formula_id}."
    return ok, "All source-backed formula dependencies are available." if ok else "One or more source-backed formula dependencies are unavailable."


def _source_fact_status(
    *,
    legacy_value: Any,
    field: Any,
    writes: Sequence[Mapping[str, Any]],
    disposition: Mapping[str, Any] | None,
) -> tuple[str, str, Any]:
    normalized_value, status, _ = _field_state(field)
    comparison = "missing_normalized_fact"
    if status == "populated" and _values_match(legacy_value, normalized_value):
        comparison = "value_match"
    elif status == "populated":
        comparison = "value_mismatch"
    accounted = bool(writes) or disposition is not None
    reproduced = comparison == "value_match" and accounted
    return ("reproduced_correctly" if reproduced else "missing_or_explicitly_unavailable", comparison, normalized_value)


def _entry(
    *,
    parity_id: str,
    domain: str,
    metric: str,
    period: str,
    dimensions: Mapping[str, Any] | None,
    legacy_range: str,
    source_kind: str,
    normalized_path: str,
    requirement: str,
    minimum: int,
    writes: Sequence[Mapping[str, Any]],
    source_ref: str = "",
    formula_cell: str = "",
    formula_present: bool | None = None,
    formula_protected: bool | None = None,
    economically_calculable: bool | None = None,
    calculation_reason: str = "",
    inventory_class: str = "source_fact",
    inventory_origin: str = "legacy_workbook_business_key",
    legacy_value: Any = None,
    normalized_value: Any = None,
    unit: str = "",
    comparison_result: str = "",
    disposition: str = "",
    current_status: str | None = None,
    rejection_reason: str = "",
) -> dict[str, Any]:
    destinations = _destinations(writes)
    if inventory_class == "formula_improvement":
        formula_contract_status = (
            "present_protected"
            if formula_present and formula_protected
            else "present_unprotected"
            if formula_present
            else "missing"
        )
        economic_status = "economically_calculable" if economically_calculable else "blank_due_to_missing_evidence"
        if formula_contract_status != "present_protected":
            formula_status = "missing_or_explicitly_unavailable"
        elif economically_calculable:
            formula_status = "reproduced_correctly"
        else:
            formula_status = "contract_present_blank_by_missing_evidence"
        reproduced = formula_status == "reproduced_correctly"
    else:
        formula_contract_status = "not_applicable"
        economic_status = "not_applicable"
        formula_status = "reproduced_correctly" if destinations else "missing_or_explicitly_unavailable"
        reproduced = bool(destinations)
    resolved_status = current_status or formula_status
    return {
        "parity_id": parity_id,
        "domain": domain,
        "metric_business_meaning": metric,
        "period": period,
        "dimensions": dict(dimensions or {}),
        "legacy_sheet_range": legacy_range,
        "source_backed_vs_derived": source_kind,
        "inventory_class": inventory_class,
        "inventory_origin": inventory_origin,
        "legacy_value": legacy_value,
        "normalized_value": normalized_value,
        "unit": unit,
        "comparison_result": comparison_result,
        "formula_contract_status": formula_contract_status,
        "economic_calculability": economic_status,
        "calculation_reason": calculation_reason,
        "disposition": disposition,
        "source_ref": source_ref,
        "normalized_package_path": normalized_path,
        "binding_ids": _binding_ids(writes),
        "formula_ownership": formula_cell or ("formula_owned" if disposition == "formula_owned" else "value_only_binding"),
        "expected_new_workbook_destination": destinations or ([formula_cell] if formula_cell else []),
        "minimum_coverage_requirement": minimum,
        "parity_requirement": requirement,
        "rejection_reason": rejection_reason or (
            "User input or unsupported legacy assumption lacks reproducible source lineage."
            if requirement == "intentionally_rejected"
            else ""
        ),
        "current_status": resolved_status,
    }


def _iso_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "").strip()


def _publication_date_from_document(document: str, fallback: Any) -> str:
    match = re.search(r"(20\d{2})[-_](\d{2})[-_](\d{2})", document)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    return _iso_value(fallback)


def _guidance_signature(item: Mapping[str, Any]) -> tuple[Any, ...]:
    value = str(_scalar(item.get("value")) or "")
    return (
        *guidance_scope_key(item),
        str(item.get("publication_date") or ""),
        re.sub(r"\s+", " ", value.strip().casefold()),
    )


def _legacy_guidance_inventory(legacy_path: Path) -> list[dict[str, Any]]:
    """Read legacy guidance business keys before consulting the package."""

    workbook = load_workbook(legacy_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Guidance_Normalized"]
        headers = {str(sheet.cell(1, column).value or ""): column for column in range(1, sheet.max_column + 1)}
        rows: list[dict[str, Any]] = []
        for row_number in range(2, sheet.max_row + 1):
            metric = str(sheet.cell(row_number, headers.get("metric", headers["metric_hint"])).value or "").strip()
            value = str(sheet.cell(row_number, headers["numbers"]).value or "").strip()
            legacy_horizon = str(sheet.cell(row_number, headers["horizon_label"]).value or "").strip()
            if not metric or not value or not legacy_horizon:
                continue
            document = str(sheet.cell(row_number, headers["doc"]).value or "").strip()
            horizon, source_table_context = _anf_legacy_guidance_horizon(
                row_number=row_number,
                document=document,
                page=sheet.cell(row_number, headers["page"]).value,
                metric=metric,
                numbers=value,
                legacy_horizon=legacy_horizon,
            )
            publication_date = _publication_date_from_document(
                document,
                sheet.cell(row_number, headers.get("source_date", headers["quarter"])).value,
            )
            source_ref = f"{document}#{legacy_path.name}!Guidance_Normalized!row:{row_number}" if document else f"{legacy_path.name}!Guidance_Normalized!row:{row_number}"
            item = {
                "metric": {"value": metric},
                "value": {"value": value, "unit": str(sheet.cell(row_number, headers["unit"]).value or "")},
                "horizon": {"value": horizon},
                "publication_date": publication_date,
                "stated_in_period": str(sheet.cell(row_number, headers["stated_in_label"]).value or "").strip(),
            }
            if source_table_context:
                item["source_table_context"] = source_table_context
            rows.append(
                {
                    "row_number": row_number,
                    "item": item,
                    "source_ref": source_ref,
                    "legacy_range": f"{legacy_path.name}!Guidance_Normalized!A{row_number}:S{row_number}",
                }
            )
        return rows
    finally:
        workbook.close()


def _guidance_parity_entries(
    package: Mapping[str, Any],
    writes_by_path: Mapping[str, Sequence[Mapping[str, Any]]],
    legacy_path: Path,
    derived_writes_by_source_ref: Mapping[
        str, Sequence[Mapping[str, Any]]
    ] | None = None,
) -> list[dict[str, Any]]:
    derived_writes_by_source_ref = derived_writes_by_source_ref or {}
    legacy_rows = _legacy_guidance_inventory(legacy_path)
    package_rows = _path_get(package, "normalized_guidance.items") or []
    package_by_signature: dict[tuple[Any, ...], list[tuple[int, Mapping[str, Any]]]] = defaultdict(list)
    for index, item in enumerate(package_rows):
        if isinstance(item, Mapping):
            package_by_signature[_guidance_signature(item)].append((index, item))

    seen_signatures: Counter[tuple[Any, ...]] = Counter()
    entries: list[dict[str, Any]] = []
    for legacy in legacy_rows:
        item = legacy["item"]
        signature = _guidance_signature(item)
        seen_signatures[signature] += 1
        matches = package_by_signature.get(signature, [])
        package_index, package_item = matches[0] if matches else (-1, {})
        path = (
            f"normalized_guidance.items.{package_index}.value"
            if package_index >= 0
            else f"normalized_guidance.items[missing:{legacy['row_number']}].value"
        )
        writes = _merge_guidance_value_writes(
            writes_by_path.get(path, []),
            package_item,
            derived_writes_by_source_ref,
        )
        evidence_refs = package_item.get("evidence_refs", []) if isinstance(package_item, Mapping) else []
        lineage_retained = legacy["source_ref"] in evidence_refs
        role = str(package_item.get("display_role") or "") if isinstance(package_item, Mapping) else ""
        routed = bool(writes) if role in {"current_primary", "current_secondary"} else role in {"history", "superseded", "audit_only"}
        reproduced = bool(matches) and lineage_retained and routed
        scope = normalize_guidance_scope(item)
        entries.append(
            _entry(
                parity_id=f"legacy-guidance:{legacy['row_number']}:{scope.metric}:{scope.horizon}",
                domain="guidance",
                metric=scope.metric,
                period=scope.horizon,
                dimensions={
                    "publication_date": item["publication_date"],
                    "source_reporting_period": item["stated_in_period"],
                    "display_role": role,
                    "duplicate_evidence": seen_signatures[signature] > 1,
                },
                legacy_range=legacy["legacy_range"],
                source_kind="source_backed",
                normalized_path=path,
                requirement="must_reproduce" if seen_signatures[signature] == 1 else "may_improve_semantically",
                minimum=1,
                writes=writes,
                source_ref=legacy["source_ref"],
                inventory_class="source_fact" if seen_signatures[signature] == 1 else "duplicate_display_use",
                inventory_origin="legacy_workbook_business_key",
                legacy_value=_scalar(item["value"]),
                normalized_value=_scalar(package_item.get("value")) if isinstance(package_item, Mapping) else None,
                unit=str(item["value"].get("unit") or ""),
                comparison_result=(
                    "semantic_value_and_lineage_match"
                    if reproduced
                    else "normalized_match_missing_exact_lineage"
                    if matches
                    else "missing_normalized_guidance"
                ),
                disposition=role or "missing",
                current_status="reproduced_correctly" if reproduced else "missing_or_explicitly_unavailable",
            )
        )
    return entries


def _legacy_narrative_scalar_entries(
    package: Mapping[str, Any],
    writes_by_path: Mapping[str, Sequence[Mapping[str, Any]]],
    legacy_path: Path,
) -> list[dict[str, Any]]:
    workbook = load_workbook(legacy_path, read_only=True, data_only=True)
    try:
        inventory = [
            {
                "parity_id": parity_id,
                "domain": domain,
                "sheet": sheet_name,
                "cell": cell,
                "legacy_value": workbook[sheet_name][cell].value,
                "path": path,
                "requirement": requirement,
            }
            for parity_id, domain, sheet_name, cell, path, requirement in LEGACY_NARRATIVE_SCALARS
        ]
    finally:
        workbook.close()

    entries: list[dict[str, Any]] = []
    for item in inventory:
        field = _path_get(package, item["path"])
        normalized_value, status, source_ref = _field_state(field)
        writes = list(writes_by_path.get(item["path"], []))
        internal_legacy = bool(re.search(r"Operating_Drivers|Investment_Case|planner|binding|parser", str(item["legacy_value"] or ""), re.I))
        reproduced = status == "populated" and bool(source_ref) and bool(writes)
        entries.append(
            _entry(
                parity_id=item["parity_id"],
                domain=item["domain"],
                metric=item["parity_id"].split(":", 1)[-1],
                period="latest",
                dimensions={},
                legacy_range=f"{legacy_path.name}!{item['sheet']}!{item['cell']}",
                source_kind="evidence_backed_synthesis",
                normalized_path=item["path"],
                requirement=item["requirement"],
                minimum=1,
                writes=writes,
                source_ref=source_ref,
                inventory_class="parser_internal_text" if internal_legacy else "source_fact",
                inventory_origin="legacy_workbook_business_key",
                legacy_value=item["legacy_value"],
                normalized_value=normalized_value,
                comparison_result="legacy_internal_text_replaced" if internal_legacy and reproduced else "investor_wording_improved" if reproduced else "missing_normalized_narrative",
                disposition="visible_improved" if reproduced else "missing",
                current_status="reproduced_with_improved_wording" if reproduced else "missing_or_explicitly_unavailable",
            )
        )
    return entries


def _find_row_path(
    package: Mapping[str, Any],
    collection_path: str,
    field_name: str,
    expected_value: str,
    value_field: str,
) -> tuple[str, Mapping[str, Any]]:
    rows = _path_get(package, collection_path)
    if not isinstance(rows, list):
        return f"{collection_path}[missing:{expected_value}].{value_field}", {}
    for index, row in enumerate(rows):
        if isinstance(row, Mapping) and str(_scalar(row.get(field_name)) or row.get(field_name) or "").strip().casefold() == expected_value.strip().casefold():
            return f"{collection_path}.{index}.{value_field}", row
    return f"{collection_path}[missing:{expected_value}].{value_field}", {}


def _legacy_narrative_row_entries(
    package: Mapping[str, Any],
    writes_by_path: Mapping[str, Sequence[Mapping[str, Any]]],
    legacy_path: Path,
) -> list[dict[str, Any]]:
    workbook = load_workbook(legacy_path, read_only=True, data_only=True)
    try:
        summary_operating = [(row, str(workbook["SUMMARY"].cell(row, 1).value or "").rstrip(":"), workbook["SUMMARY"].cell(row, 2).value) for row in range(13, 16)]
        summary_dependencies = [(row, workbook["SUMMARY"].cell(row, 1).value) for row in range(17, 22)]
        summary_invalidators = [(row, workbook["SUMMARY"].cell(row, 1).value) for row in range(23, 25)]
        driver_topics = [(row, str(workbook["Operating_Drivers"].cell(row, 1).value or "")) for row in range(6, 10)]
        quarter_note_rows = [(row, workbook["Quarter_Notes_UI"].cell(row, 3).value) for row in range(10, 16)]
    finally:
        workbook.close()

    entries: list[dict[str, Any]] = []

    for row_number, member, legacy_text in summary_operating:
        path, row = _find_row_path(package, "company_profile.operating_model_rows", "member", member, "description")
        writes = list(writes_by_path.get(path, []))
        field = row.get("description") if isinstance(row, Mapping) else None
        value, status, source_ref = _field_state(field)
        entries.append(_entry(
            parity_id=f"summary:operating-model:{member}", domain="summary", metric="operating model by geography",
            period="latest", dimensions={"member": member}, legacy_range=f"{legacy_path.name}!SUMMARY!A{row_number}:B{row_number}",
            source_kind="evidence_backed_synthesis", normalized_path=path, requirement="may_improve_semantically", minimum=1,
            writes=writes, source_ref=source_ref, legacy_value=legacy_text, normalized_value=value,
            comparison_result="investor_wording_improved" if status == "populated" and writes else "missing_normalized_narrative",
            disposition="visible_improved" if writes else "missing", current_status="reproduced_with_improved_wording" if writes else "missing_or_explicitly_unavailable",
        ))

    dependency_rows = _path_get(package, "company_profile.key_dependencies") or []
    for offset, (row_number, legacy_text) in enumerate(summary_dependencies):
        package_row = dependency_rows[offset] if offset < len(dependency_rows) and isinstance(dependency_rows[offset], Mapping) else {}
        path = f"company_profile.key_dependencies.{offset}.text" if package_row else f"company_profile.key_dependencies[missing:{row_number}].text"
        field = package_row.get("text") if package_row else None
        value, status, source_ref = _field_state(field)
        writes = list(writes_by_path.get(path, []))
        entries.append(_entry(
            parity_id=f"summary:dependency:{row_number}", domain="summary", metric="key dependency", period="latest", dimensions={"display_order": offset + 1},
            legacy_range=f"{legacy_path.name}!SUMMARY!A{row_number}", source_kind="evidence_backed_synthesis", normalized_path=path,
            requirement="may_improve_semantically", minimum=1, writes=writes, source_ref=source_ref,
            legacy_value=legacy_text, normalized_value=value, comparison_result="investor_wording_improved" if status == "populated" and writes else "missing_normalized_narrative",
            disposition="visible_improved" if writes else "missing", current_status="reproduced_with_improved_wording" if writes else "missing_or_explicitly_unavailable",
        ))

    invalidator_rows = _path_get(package, "investment_case.invalidators") or []
    for offset, (row_number, legacy_text) in enumerate(summary_invalidators):
        package_row = invalidator_rows[offset] if offset < len(invalidator_rows) and isinstance(invalidator_rows[offset], Mapping) else {}
        path = f"investment_case.invalidators.{offset}.text" if package_row else f"investment_case.invalidators[missing:{row_number}].text"
        field = package_row.get("text") if package_row else None
        value, status, source_ref = _field_state(field)
        writes = list(writes_by_path.get(path, []))
        entries.append(_entry(
            parity_id=f"summary:invalidator:{row_number}", domain="summary", metric="thesis invalidator", period="latest", dimensions={"display_order": offset + 1},
            legacy_range=f"{legacy_path.name}!SUMMARY!A{row_number}", source_kind="analyst_interpretation_requiring_review", normalized_path=path,
            requirement="may_improve_semantically", minimum=1, writes=writes, source_ref=source_ref,
            legacy_value=legacy_text, normalized_value=value, comparison_result="unsupported_legacy_generalization_replaced" if status == "populated" and writes else "missing_normalized_narrative",
            disposition="visible_reviewed_interpretation" if writes else "missing", current_status="reproduced_with_improved_wording" if writes else "missing_or_explicitly_unavailable",
        ))

    driver_aliases = {"Sales guide": "Sales execution", "Margin durability": "Margin durability", "Inventory quality": "Inventory quality", "Capital returns": "Capital returns"}
    for row_number, legacy_topic in driver_topics:
        expected_topic = driver_aliases.get(legacy_topic, legacy_topic)
        path, row = _find_row_path(package, "operating_drivers.items", "topic", expected_topic, "current_read")
        field = row.get("current_read") if isinstance(row, Mapping) else None
        value, status, source_ref = _field_state(field)
        writes = list(writes_by_path.get(path, []))
        entries.append(_entry(
            parity_id=f"driver:{legacy_topic.casefold().replace(' ', '-')}", domain="operating_drivers", metric=legacy_topic,
            period=str(row.get("period") or "latest") if isinstance(row, Mapping) else "latest", dimensions={},
            legacy_range=f"{legacy_path.name}!Operating_Drivers!A{row_number}:H{row_number}", source_kind="evidence_backed_synthesis",
            normalized_path=path, requirement="must_reproduce", minimum=1, writes=writes, source_ref=source_ref,
            normalized_value=value, comparison_result="source_backed_theme_improved" if status == "populated" and writes else "missing_normalized_driver",
            disposition="visible_improved" if writes else "missing", current_status="reproduced_with_improved_wording" if writes else "missing_or_explicitly_unavailable",
        ))

    note_mapping = {10: "Q4 results", 11: "2026 margin bridge"}
    rejected_reasons = {
        12: "Legacy row contains implementation language and duplicates current guidance.",
        13: "Legacy revolver-change row is duplicated liquidity content and is not a clean quarter theme.",
        14: "Legacy row duplicates operating-margin guidance already routed to guidance blocks.",
        15: "Legacy row duplicates revenue guidance already routed to guidance blocks.",
    }
    for row_number, legacy_text in quarter_note_rows:
        if row_number in rejected_reasons:
            entries.append(_entry(
                parity_id=f"quarter-note:legacy-row-{row_number}", domain="quarter_notes", metric="legacy quarter-note candidate", period="2026-Q1", dimensions={},
                legacy_range=f"{legacy_path.name}!Quarter_Notes_UI!A{row_number}:M{row_number}", source_kind="legacy_only",
                normalized_path="", requirement="intentionally_rejected", minimum=0, writes=[], inventory_class="parser_internal_text" if row_number == 12 else "duplicate_display_use",
                inventory_origin="legacy_workbook_business_key", legacy_value=legacy_text, comparison_result="intentionally_rejected",
                disposition="rejected", current_status="reproduced_correctly", rejection_reason=rejected_reasons[row_number],
            ))
            continue
        theme = note_mapping[row_number]
        path, row = _find_row_path(package, "quarter_notes.items", "theme", theme, "commentary")
        field = row.get("commentary") if isinstance(row, Mapping) else None
        value, status, source_ref = _field_state(field)
        writes = list(writes_by_path.get(path, []))
        entries.append(_entry(
            parity_id=f"quarter-note:legacy-row-{row_number}", domain="quarter_notes", metric=theme, period=str(_scalar(row.get("quarter")) or "2025-Q4") if isinstance(row, Mapping) else "2025-Q4", dimensions={},
            legacy_range=f"{legacy_path.name}!Quarter_Notes_UI!A{row_number}:M{row_number}", source_kind="evidence_backed_synthesis",
            normalized_path=path, requirement="must_reproduce", minimum=1, writes=writes, source_ref=source_ref,
            legacy_value=legacy_text, normalized_value=value, inventory_class="parser_internal_text", comparison_result="legacy_internal_text_replaced" if status == "populated" and writes else "missing_clean_quarter_note",
            disposition="visible_improved" if writes else "missing", current_status="reproduced_with_improved_wording" if writes else "missing_or_explicitly_unavailable",
        ))

    for theme in ("Brand mix", "Inventory", "Capital allocation", "Growth channels"):
        path, row = _find_row_path(package, "quarter_notes.items", "theme", theme, "commentary")
        field = row.get("commentary") if isinstance(row, Mapping) else None
        value, status, source_ref = _field_state(field)
        writes = list(writes_by_path.get(path, []))
        entries.append(_entry(
            parity_id=f"quarter-note:source:{theme.casefold().replace(' ', '-')}", domain="quarter_notes", metric=theme,
            period=str(_scalar(row.get("quarter")) or "2025-Q4") if isinstance(row, Mapping) else "2025-Q4", dimensions={},
            legacy_range=source_ref, source_kind="source_backed", normalized_path=path, requirement="must_reproduce", minimum=1,
            writes=writes, source_ref=source_ref, normalized_value=value, inventory_origin="source_evidence_business_key",
            comparison_result="source_backed_theme_added" if status == "populated" and writes else "missing_clean_quarter_note",
            disposition="visible_improved" if writes else "missing", current_status="reproduced_correctly" if writes else "missing_or_explicitly_unavailable",
        ))
    return entries


def _promise_progress_parity_entries(
    package: Mapping[str, Any],
    writes_by_path: Mapping[str, Sequence[Mapping[str, Any]]],
    legacy_path: Path,
    derived_writes_by_source_ref: Mapping[
        str, Sequence[Mapping[str, Any]]
    ] | None = None,
) -> list[dict[str, Any]]:
    """Inventory legacy Promise Progress scopes before matching normalized routes."""

    derived_writes_by_source_ref = derived_writes_by_source_ref or {}
    legacy_metric_aliases = {
        "eps": "Adj EPS",
        "diluted-share": "Diluted shares",
        "real-estate": "Real estate activity",
        "tariff-impact": "Tariffs",
    }
    workbook = load_workbook(legacy_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Promise_Progress"]
        headers = {str(sheet.cell(1, column).value or ""): column for column in range(1, sheet.max_column + 1)}
        grouped: dict[tuple[str, int], dict[str, Any]] = {}
        for row_number in range(2, sheet.max_row + 1):
            metric_text = str(sheet.cell(row_number, headers["metric_display"]).value or sheet.cell(row_number, headers["metric_ref"]).value or "").strip()
            horizon = str(sheet.cell(row_number, headers["target_period_label"]).value or "").strip()
            if not metric_text or not horizon:
                continue
            metric_text = re.sub(r"\s+guidance$", "", metric_text, flags=re.I)
            metric_text = legacy_metric_aliases.get(metric_text.casefold(), metric_text)
            scope = normalize_guidance_scope({"metric": {"value": metric_text}, "horizon": {"value": horizon}, "value": {"unit": ""}})
            if scope.fiscal_year is None:
                continue
            key = (scope.metric, int(scope.fiscal_year))
            group = grouped.setdefault(key, {"rows": [], "values": [], "horizon": scope.horizon})
            group["rows"].append(row_number)
            target = sheet.cell(row_number, headers["target_display"]).value
            if target not in (None, ""):
                group["values"].append(str(target))
    finally:
        workbook.close()

    promise_rows = _path_get(package, "promise_progress.items") or []
    promise_index: dict[tuple[str, int], tuple[int, Mapping[str, Any]]] = {}
    for index, row in enumerate(promise_rows):
        if not isinstance(row, Mapping):
            continue
        scope = normalize_guidance_scope({"metric": row.get("metric"), "horizon": row.get("horizon"), "value": row.get("current_guidance")})
        if scope.fiscal_year is not None:
            promise_index[(scope.metric, int(scope.fiscal_year))] = (index, row)

    guidance_rows = _path_get(package, "normalized_guidance.items") or []
    guidance_index: dict[tuple[str, int], list[tuple[int, Mapping[str, Any]]]] = defaultdict(list)
    for index, row in enumerate(guidance_rows):
        if not isinstance(row, Mapping):
            continue
        scope = normalize_guidance_scope(row)
        if scope.fiscal_year is not None and scope.horizon_type == "FY":
            guidance_index[(scope.metric, int(scope.fiscal_year))].append((index, row))

    historical_rows = _path_get(package, "promise_progress.historical_evidence_items") or []
    historical_index: dict[tuple[str, int], list[tuple[int, Mapping[str, Any]]]] = defaultdict(list)
    for index, row in enumerate(historical_rows):
        if not isinstance(row, Mapping):
            continue
        scope = normalize_guidance_scope(
            {
                "metric": {"value": row.get("metric")},
                "horizon": {"value": row.get("horizon")},
                "value": {"unit": ""},
            }
        )
        if scope.fiscal_year is not None and scope.horizon_type == "FY":
            historical_index[(scope.metric, int(scope.fiscal_year))].append((index, row))

    entries: list[dict[str, Any]] = []
    for (metric, fiscal_year), inventory in sorted(grouped.items()):
        promise_match = promise_index.get((metric, fiscal_year))
        if promise_match:
            index, row = promise_match
            path = f"promise_progress.items.{index}.current_guidance"
            item_prefix = f"promise_progress.items.{index}."
            writes = [
                write
                for write_path, path_writes in writes_by_path.items()
                if str(write_path).startswith(item_prefix)
                for write in path_writes
            ]
            role = str(row.get("display_role") or "")
            disposition = str(row.get("visibility_disposition") or role)
            normalized_value = _scalar(row.get("current_guidance"))
            parity_category = "visible_reproduced" if writes else "audit_only" if disposition == "audit_only" else "historical_reproduced"
            current_status = "reproduced_correctly"
            rejection_reason = ""
            disposition_paths: list[str] = []
            disposition_source_refs: list[str] = []
            occurrence_dispositions: Counter[str] = Counter()
        else:
            candidates = sorted(
                guidance_index.get((metric, fiscal_year), []),
                key=lambda pair: (str(pair[1].get("publication_date") or ""), str(pair[1].get("evidence_key") or "")),
            )
            if candidates:
                index, row = candidates[-1]
                path = f"normalized_guidance.items.{index}.value"
                role = str(row.get("display_role") or "")
                disposition = "current_guidance_visible_elsewhere" if role in {"current_primary", "current_secondary"} else role
                normalized_value = _scalar(row.get("value"))
                parity_category = (
                    "visible_reproduced"
                    if disposition == "current_guidance_visible_elsewhere"
                    else "historical_reproduced"
                    if role == "history"
                    else "duplicate_superseded"
                    if role == "superseded"
                    else "audit_only"
                )
                current_status = "reproduced_correctly"
                rejection_reason = ""
                disposition_paths = []
                disposition_source_refs = []
                occurrence_dispositions = Counter()
            else:
                disposition_rows = historical_index.get((metric, fiscal_year), [])
                if disposition_rows:
                    indexes = [index for index, _row in disposition_rows]
                    rows_with_disposition = [row for _index, row in disposition_rows]
                    occurrence_dispositions = Counter(str(row.get("disposition") or "") for row in rows_with_disposition)
                    disposition_paths = [f"promise_progress.historical_evidence_items.{index}" for index in indexes]
                    disposition_source_refs = list(
                        dict.fromkeys(
                            ref
                            for row in rows_with_disposition
                            for ref in (row.get("source_refs") or [])
                            if ref
                        )
                    )
                    if occurrence_dispositions["audit_only_historical_evidence"]:
                        parity_category = "audit_only"
                        disposition = "audit_only_historical_evidence"
                        current_status = "audit_only_evidence_preserved"
                    elif occurrence_dispositions["duplicate_or_superseded_evidence"]:
                        parity_category = "duplicate_superseded"
                        disposition = "duplicate_or_superseded_evidence"
                        current_status = "duplicate_or_superseded_evidence_preserved"
                    elif occurrence_dispositions["rejected_with_evidence"]:
                        parity_category = "rejected_with_evidence"
                        disposition = "rejected_with_evidence"
                        current_status = "explicitly_rejected_with_evidence"
                    else:
                        parity_category = "unavailable_without_adequate_evidence"
                        disposition = "unavailable_without_adequate_evidence"
                        current_status = "unavailable_without_adequate_evidence"
                    path = disposition_paths[0]
                    row = rows_with_disposition[0]
                    role = ""
                    normalized_value = [str(item.get("target_value") or "") for item in rows_with_disposition]
                    rejection_reason = "; ".join(
                        dict.fromkeys(
                            str(item.get("disposition_reason") or "")
                            for item in rows_with_disposition
                            if item.get("disposition") == "rejected_with_evidence"
                        )
                    )
                else:
                    path = f"promise_progress.items[missing:{metric}:FY{fiscal_year}].current_guidance"
                    row = {}
                    role = ""
                    disposition = "unavailable_without_adequate_evidence"
                    normalized_value = None
                    parity_category = "unavailable_without_adequate_evidence"
                    current_status = "unavailable_without_adequate_evidence"
                    rejection_reason = "No normalized route or evidence-preserving disposition exists for the legacy business key."
                    disposition_paths = []
                    disposition_source_refs = []
                    occurrence_dispositions = Counter()
            writes = _merge_guidance_value_writes(
                writes_by_path.get(path, []),
                row,
                derived_writes_by_source_ref,
            )
        rows = inventory["rows"]
        entries.append(_entry(
            parity_id=f"promise-progress:{metric}:FY{fiscal_year}", domain="promise_progress", metric=metric,
            period=f"FY{fiscal_year}", dimensions={
                "legacy_occurrence_count": len(rows),
                "route": disposition,
                "promise_parity_category": parity_category,
                "disposition_paths": disposition_paths,
                "source_refs": disposition_source_refs,
                "occurrence_disposition_counts": dict(sorted(occurrence_dispositions.items())),
            },
            legacy_range=f"{legacy_path.name}!Promise_Progress!A{min(rows)}:V{max(rows)}", source_kind="derived_from_guidance_evidence",
            normalized_path=path, requirement="may_improve_semantically", minimum=1, writes=writes,
            source_ref=(
                disposition_source_refs[0]
                if disposition_source_refs
                else str((row.get("evidence_refs") or [""])[0])
                if isinstance(row, Mapping) and row.get("evidence_refs")
                else str(row.get("source_ref") or "")
                if isinstance(row, Mapping)
                else ""
            ),
            inventory_class="duplicate_display_use", inventory_origin="legacy_workbook_business_key",
            legacy_value=sorted(set(inventory["values"])), normalized_value=normalized_value,
            comparison_result=(
                "routed_to_progression_or_guidance"
                if current_status == "reproduced_correctly"
                else "evidence_preserved_with_explicit_disposition"
                if current_status != "unavailable_without_adequate_evidence"
                else "unavailable_without_adequate_evidence"
            ),
            disposition=disposition,
            current_status=current_status,
            rejection_reason=rejection_reason,
        ))
    return entries


def build_parity_matrix(
    *,
    package: Mapping[str, Any],
    plan: Mapping[str, Any],
    legacy_path: Path,
    shell_path: Path,
    binding_path: Path,
) -> dict[str, Any]:
    if str(plan.get("status") or "") != "PASS":
        raise ValueError("Parity matrix requires the current blocker-free PASS plan.")
    writes_by_path = _write_index(plan)
    entries: list[dict[str, Any]] = []

    binding_document = load_json_strict(binding_path)
    guidance_destination_lineage = _guidance_destination_lineage(plan, binding_document)
    derived_guidance_writes = _derived_guidance_value_write_index(
        plan,
        binding_document,
    )
    dispositions = _financial_dispositions(binding_document)
    selector_exclusions = _selector_exclusion_index(plan)
    history_rows, history_headers = _legacy_history_rows(legacy_path)
    unsupported_zero_placeholders = _legacy_unsupported_zero_placeholders(legacy_path, history_rows)
    quarterly_package_rows = _package_row_index(package, "quarterly_financials")
    annual_package_rows = _package_row_index(package, "annual_financials")
    history_period_ordinals, history_values = _history_evidence_index(package)

    # Inventory starts with the latest twelve legacy business periods, not the
    # package.  A missing package row therefore remains visible as a parity gap.
    for legacy_row in history_rows[-12:]:
        period = str(legacy_row["period"])
        package_row_info = quarterly_package_rows.get(period)
        for legacy_header, (metric, unit, domain) in LEGACY_HISTORY_METRICS.items():
            column = history_headers[legacy_header]
            history_ref = f"{legacy_path.name}!History_Q!{get_column_letter(column)}{legacy_row['row_number']}"
            source_value, source_ref = _anf_history_source_evidence(
                legacy_row["values"],
                legacy_header,
                history_ref,
            )
            legacy_value = _normalized_legacy_value(source_value, unit)
            if legacy_value is None:
                continue
            if package_row_info is None:
                index, package_row = -1, {}
                path = f"quarterly_financials.rows[missing:{period}].{metric}"
            else:
                index, package_row = package_row_info
                path = f"quarterly_financials.rows.{index}.{metric}"
            writes = writes_by_path.get(path, [])
            disposition, disposition_name = _effective_disposition(
                static_disposition=dispositions.get(("quarterly_financials", metric)),
                selector_exclusion=selector_exclusions.get(path),
                writes=writes,
            )
            placeholder = unsupported_zero_placeholders.get((metric, period))
            if placeholder is not None:
                normalized_value, normalized_status, _ = _field_state(package_row.get(metric))
                report_refs = " + ".join(
                    filter(
                        None,
                        (
                            str(placeholder.get("value_source_ref") or ""),
                            str(placeholder.get("metadata_source_ref") or ""),
                        ),
                    )
                )
                entries.append(
                    _entry(
                        parity_id=f"legacy-quarter:{period}:{metric}",
                        domain=domain,
                        metric=metric,
                        period=period,
                        dimensions={},
                        legacy_range=f"{history_ref} + {report_refs}",
                        source_kind="unavailable",
                        normalized_path=path,
                        requirement="unavailable_missing_evidence",
                        minimum=1,
                        writes=writes,
                        source_ref=report_refs,
                        inventory_class="unsupported_legacy_content",
                        inventory_origin="legacy_report_quality_check",
                        legacy_value=None,
                        normalized_value=normalized_value,
                        unit=unit,
                        comparison_result=(
                            "unsupported_zero_placeholder_left_blank"
                            if normalized_status != "populated"
                            else "unsupported_zero_placeholder_incorrectly_populated"
                        ),
                        disposition="leave_blank",
                        current_status="missing_or_explicitly_unavailable",
                        rejection_reason=(
                            "The legacy report marks this zero candidate Source=Missing or QA=FAIL; "
                            "it is a placeholder, not source-backed financial evidence."
                        ),
                    )
                )
                continue
            status, comparison, normalized_value = _source_fact_status(
                legacy_value=legacy_value,
                field=package_row.get(metric),
                writes=writes,
                disposition=disposition,
            )
            derived_metric = legacy_header in LEGACY_DERIVED_HISTORY_METRICS
            entries.append(
                _entry(
                    parity_id=f"legacy-quarter:{period}:{metric}",
                    domain=domain,
                    metric=metric,
                    period=period,
                    dimensions={},
                    legacy_range=source_ref,
                    source_kind="derived" if derived_metric else "source_backed",
                    normalized_path=path,
                    requirement="must_reproduce",
                    minimum=1,
                    writes=writes,
                    source_ref=source_ref,
                    inventory_class="source_fact",
                    legacy_value=legacy_value,
                    normalized_value=normalized_value,
                    unit=unit,
                    comparison_result=comparison,
                    disposition=disposition_name,
                    current_status=status,
                )
            )

    # Adjusted EBITDA and adjusted EPS are visible legacy definitions outside
    # History_Q.  They are independently inventoried from the Valuation matrix.
    for legacy_row_number, metric, unit in ((24, "adjusted_ebitda", "$m"), (110, "adjusted_eps", "$/share")):
        for period, (legacy_value, source_ref) in _legacy_valuation_series(legacy_path, legacy_row_number).items():
            package_row_info = quarterly_package_rows.get(period)
            if package_row_info is None:
                path, package_row = f"quarterly_financials.rows[missing:{period}].{metric}", {}
            else:
                index, package_row = package_row_info
                path = f"quarterly_financials.rows.{index}.{metric}"
            writes = writes_by_path.get(path, [])
            status, comparison, normalized_value = _source_fact_status(
                legacy_value=round(legacy_value, 6),
                field=package_row.get(metric),
                writes=writes,
                disposition=dispositions.get(("quarterly_financials", metric)),
            )
            entries.append(
                _entry(
                    parity_id=f"legacy-quarter:{period}:{metric}", domain="per_share" if metric.endswith("eps") else "quarterly_financials",
                    metric=metric, period=period, dimensions={}, legacy_range=source_ref,
                    source_kind="source_backed", normalized_path=path, requirement="must_reproduce",
                    minimum=1, writes=writes, source_ref=source_ref,
                    legacy_value=round(legacy_value, 6), normalized_value=normalized_value, unit=unit,
                    comparison_result=comparison, disposition="planned_binding" if writes else "", current_status=status,
                )
            )

    # Annual source facts are independently reconstructed only from complete
    # fiscal Q1-Q4 sets.  Annual EPS and weighted-average shares are deliberately
    # absent unless an annual source supplies them.
    by_year: dict[int, dict[int, Mapping[str, Any]]] = defaultdict(dict)
    for row in history_rows:
        by_year[int(row["fiscal_year"])][int(row["fiscal_quarter"])] = row
    adjusted_by_period = _legacy_valuation_series(legacy_path, 24)
    for year in sorted(by_year):
        quarters = by_year[year]
        period = f"{year}-FY"
        if set(quarters) != {1, 2, 3, 4}:
            continue
        package_row_info = annual_package_rows.get(period)
        if package_row_info is None:
            index, package_row = -1, {}
        else:
            index, package_row = package_row_info
        header_by_metric = {metric: header for header, (metric, _unit, _domain) in LEGACY_HISTORY_METRICS.items()}
        annual_fields = sorted(ANNUAL_FLOW_FIELDS | ANNUAL_POINT_IN_TIME_FIELDS)
        for metric in annual_fields:
            legacy_header = header_by_metric.get(metric)
            if not legacy_header:
                continue
            unit = LEGACY_HISTORY_METRICS[legacy_header][1]
            component_rows = [quarters[quarter] for quarter in (1, 2, 3, 4)]
            placeholder_rows = component_rows if metric in ANNUAL_FLOW_FIELDS else [component_rows[-1]]
            placeholder_details = [
                unsupported_zero_placeholders[(metric, str(row["period"]))]
                for row in placeholder_rows
                if (metric, str(row["period"])) in unsupported_zero_placeholders
            ]
            if placeholder_details:
                path = (
                    f"annual_financials.rows.{index}.{metric}"
                    if package_row_info is not None
                    else f"annual_financials.rows[missing:{period}].{metric}"
                )
                writes = writes_by_path.get(path, [])
                normalized_value, normalized_status, _ = _field_state(package_row.get(metric))
                report_refs = " + ".join(
                    sorted(
                        {
                            str(detail.get(ref_name) or "")
                            for detail in placeholder_details
                            for ref_name in ("value_source_ref", "metadata_source_ref")
                            if detail.get(ref_name)
                        }
                    )
                )
                entries.append(
                    _entry(
                        parity_id=f"legacy-annual:{period}:{metric}",
                        domain=(
                            "annual_financials"
                            if metric not in ANNUAL_POINT_IN_TIME_FIELDS
                            else "balance_sheet"
                        ),
                        metric=metric,
                        period=period,
                        dimensions={},
                        legacy_range=report_refs,
                        source_kind="unavailable",
                        normalized_path=path,
                        requirement="unavailable_missing_evidence",
                        minimum=1,
                        writes=writes,
                        source_ref=report_refs,
                        inventory_class="unsupported_legacy_content",
                        inventory_origin="legacy_report_quality_check",
                        legacy_value=None,
                        normalized_value=normalized_value,
                        unit=unit,
                        comparison_result=(
                            "unsupported_zero_placeholder_left_blank"
                            if normalized_status != "populated"
                            else "unsupported_zero_placeholder_incorrectly_populated"
                        ),
                        disposition="leave_blank",
                        current_status="missing_or_explicitly_unavailable",
                        rejection_reason=(
                            "One or more annual components are legacy zero placeholders marked "
                            "Source=Missing or QA=FAIL; no annual value may be derived."
                        ),
                    )
                )
                continue
            component_evidence = [
                _anf_history_source_evidence(
                    row["values"],
                    legacy_header,
                    f"{legacy_path.name}!History_Q!{get_column_letter(history_headers[legacy_header])}{row['row_number']}",
                )
                for row in component_rows
            ]
            component_values = [_normalized_legacy_value(value, unit) for value, _ in component_evidence]
            if metric in ANNUAL_FLOW_FIELDS:
                if any(value is None for value in component_values):
                    continue
                legacy_value = round(sum(float(value) for value in component_values), 6)
            else:
                legacy_value = component_values[-1]
                if legacy_value is None:
                    continue
            path = (
                f"annual_financials.rows.{index}.{metric}"
                if package_row_info is not None
                else f"annual_financials.rows[missing:{period}].{metric}"
            )
            writes = writes_by_path.get(path, [])
            disposition, disposition_name = _effective_disposition(
                static_disposition=dispositions.get(("annual_financials", metric)),
                selector_exclusion=selector_exclusions.get(path),
                writes=writes,
            )
            status, comparison, normalized_value = _source_fact_status(
                legacy_value=legacy_value, field=package_row.get(metric), writes=writes, disposition=disposition,
            )
            source_refs = (
                [source_ref for _, source_ref in component_evidence]
                if metric in ANNUAL_FLOW_FIELDS
                else [component_evidence[-1][1]]
            )
            entries.append(
                _entry(
                    parity_id=f"legacy-annual:{period}:{metric}", domain="annual_financials" if metric not in ANNUAL_POINT_IN_TIME_FIELDS else "balance_sheet",
                    metric=metric, period=period, dimensions={}, legacy_range=" + ".join(source_refs),
                    source_kind="source_backed", normalized_path=path, requirement="must_reproduce", minimum=1,
                    writes=writes, source_ref=" + ".join(source_refs), legacy_value=legacy_value,
                    normalized_value=normalized_value, unit=unit, comparison_result=comparison,
                    disposition=disposition_name, current_status=status,
                )
            )
        adjusted_components = [adjusted_by_period.get(str(quarters[q]["period"])) for q in (1, 2, 3, 4)]
        if all(component is not None for component in adjusted_components):
            legacy_value = round(sum(float(component[0]) for component in adjusted_components if component), 6)
            metric = "adjusted_ebitda"
            path = (
                f"annual_financials.rows.{index}.{metric}"
                if package_row_info is not None
                else f"annual_financials.rows[missing:{period}].{metric}"
            )
            writes = writes_by_path.get(path, [])
            disposition, disposition_name = _effective_disposition(
                static_disposition=dispositions.get(("annual_financials", metric)),
                selector_exclusion=selector_exclusions.get(path),
                writes=writes,
            )
            status, comparison, normalized_value = _source_fact_status(
                legacy_value=legacy_value, field=package_row.get(metric), writes=writes,
                disposition=disposition,
            )
            source_refs = [str(component[1]) for component in adjusted_components if component]
            entries.append(_entry(
                parity_id=f"legacy-annual:{period}:{metric}", domain="annual_financials", metric=metric,
                period=period, dimensions={}, legacy_range=" + ".join(source_refs), source_kind="source_backed",
                normalized_path=path, requirement="must_reproduce", minimum=1, writes=writes,
                source_ref=" + ".join(source_refs), legacy_value=legacy_value, normalized_value=normalized_value,
                unit="$m", comparison_result=comparison, disposition=disposition_name, current_status=status,
            ))
        for unavailable_metric in ("diluted_shares", "eps"):
            path = (
                f"annual_financials.rows.{index}.{unavailable_metric}"
                if package_row_info is not None
                else f"annual_financials.rows[missing:{period}].{unavailable_metric}"
            )
            entries.append(_entry(
                parity_id=f"legacy-annual:{period}:{unavailable_metric}:unsupported-proxy", domain="per_share",
                metric=unavailable_metric, period=period, dimensions={}, legacy_range="",
                source_kind="unavailable", normalized_path=path, requirement="unavailable_missing_evidence",
                minimum=1, writes=writes_by_path.get(path, []), inventory_class="unsupported_legacy_content",
                inventory_origin="independent_legacy_absence_check", comparison_result="no_source_backed_annual_denominator_or_eps",
                disposition="leave_blank", current_status="missing_or_explicitly_unavailable",
                rejection_reason="The legacy fixture has no source-backed annual diluted EPS or annual weighted-average diluted-share denominator; fiscal-Q4 shares are not a valid proxy.",
            ))

    # Segment inventory is read directly from the visible legacy matrix.  The
    # normalized package and plan are only queried after each exact cell key is known.
    segment_package_rows = _path_get(package, "segments.items") or []
    segment_index: dict[tuple[str, str, str, str], tuple[int, Mapping[str, Any]]] = {}
    for index, row in enumerate(segment_package_rows):
        if isinstance(row, Mapping):
            field_name = "annual_revenue" if str(row.get("period_type") or "") == "annual" else "revenue"
            segment_index.setdefault((str(row.get("dimension") or ""), str(row.get("member") or ""), str(row.get("period") or ""), field_name), (index, row))
    legacy_wb = load_workbook(legacy_path, read_only=True, data_only=True)
    try:
        ws = legacy_wb["BS_Segments"]
        for period_type, header_row, member_rows, columns in (
            ("quarterly", 7, (61, 62, 63, 65, 66, 67), range(2, 14)),
            ("annual", 70, (72, 73, 74), range(2, 10)),
        ):
            for row_number in member_rows:
                member = str(ws.cell(row_number, 1).value or "").strip()
                dimension = "geography" if member in {"Americas", "EMEA", "APAC"} else "brand" if member in {"Hollister", "Abercrombie"} else "total_company"
                for column in columns:
                    raw_period, raw_value = ws.cell(header_row, column).value, ws.cell(row_number, column).value
                    if raw_period in (None, "") or not isinstance(raw_value, (int, float)) or isinstance(raw_value, bool):
                        continue
                    period = f"{int(raw_period)}-FY" if period_type == "annual" else str(raw_period)
                    field_name = "annual_revenue" if period_type == "annual" else "revenue"
                    package_info = segment_index.get((dimension, member, period, field_name))
                    if package_info is None:
                        path, package_row = f"segments.items[missing:{dimension}:{member}:{period}].{field_name}", {}
                    else:
                        index, package_row = package_info
                        path = f"segments.items.{index}.{field_name}"
                    legacy_value = float(raw_value)
                    writes = writes_by_path.get(path, [])
                    status, comparison, normalized_value = _source_fact_status(
                        legacy_value=legacy_value, field=package_row.get(field_name), writes=writes, disposition=None,
                    )
                    source_ref = f"{legacy_path.name}!BS_Segments!{get_column_letter(column)}{row_number}"
                    entries.append(_entry(
                        parity_id=f"legacy-segment:{period}:{dimension}:{member}:revenue", domain="segments",
                        metric="segment revenue", period=period, dimensions={"dimension": dimension, "member": member},
                        legacy_range=source_ref, source_kind="source_backed", normalized_path=path,
                        requirement="must_reproduce", minimum=1, writes=writes, source_ref=source_ref,
                        legacy_value=legacy_value, normalized_value=normalized_value, unit="$m",
                        comparison_result=comparison, disposition="planned_binding" if writes else "", current_status=status,
                    ))
    finally:
        legacy_wb.close()

    entries.extend(
        _guidance_parity_entries(
            package,
            writes_by_path,
            legacy_path,
            derived_guidance_writes,
        )
    )
    entries.extend(
        _promise_progress_parity_entries(
            package,
            writes_by_path,
            legacy_path,
            derived_guidance_writes,
        )
    )
    entries.extend(_legacy_narrative_scalar_entries(package, writes_by_path, legacy_path))
    entries.extend(_legacy_narrative_row_entries(package, writes_by_path, legacy_path))

    for domain, path, requirement in SCALAR_REQUIREMENTS:
        value, status, source_ref = _field_state(_path_get(package, path))
        effective = requirement
        retired_duplicate = path in RETIRED_DUPLICATE_VALUATION_INPUT_PATHS
        inventory_class = "duplicate_display_use" if domain == "valuation_inputs" else "source_fact"
        inventory_origin = (
            "legacy_visible_display_contract"
            if domain == "valuation_inputs"
            else "deferred_pass_2_package_projection"
        )
        entries.append(
            _entry(
                parity_id=f"scalar:{path}",
                domain=domain,
                metric=path.rsplit(".", 1)[-1],
                period=str((_path_get(package, path) or {}).get("period") or "latest") if isinstance(_path_get(package, path), Mapping) else "latest",
                dimensions={},
                legacy_range=source_ref,
                source_kind="source_backed" if status == "populated" else "unavailable",
                normalized_path=path,
                requirement=effective,
                minimum=1,
                writes=writes_by_path.get(path, []),
                source_ref=source_ref,
                inventory_class=inventory_class,
                inventory_origin=inventory_origin,
                disposition="duplicate_display_binding" if domain == "valuation_inputs" and status == "populated" else "",
                current_status="explicitly_rejected_with_evidence" if retired_duplicate else None,
                rejection_reason=(
                    "B2 retired the duplicate Valuation display binding; the accepted canonical "
                    "quarterly-series and Investment Case formula owners remain authoritative."
                    if retired_duplicate
                    else ""
                ),
            )
        )

    wb = load_workbook(shell_path, read_only=False, data_only=False)
    legacy = load_workbook(legacy_path, read_only=False, data_only=False)
    try:
        axes = plan.get("period_axes") or {}
        for contracts, axis_id in ((FORMULA_ROWS, "valuation_quarterly_periods"),):
            axis = axes.get(axis_id) if isinstance(axes, Mapping) else {}
            period_to_column = axis.get("period_to_column") if isinstance(axis, Mapping) else {}
            for contract in contracts:
                for period, column in sorted((period_to_column or {}).items()):
                    coordinate = f"{column}{contract.row}"
                    cell = wb[contract.sheet][coordinate]
                    formula = cell.value
                    legacy_range = f"{contract.sheet}!{column}{contract.row}"
                    formula_present = isinstance(formula, str) and formula.startswith("=")
                    formula_protected = bool(cell.protection.locked)
                    calculable, calculation_reason = _quarter_formula_calculability(
                        contract.formula_id,
                        str(period),
                        period_ordinals=history_period_ordinals,
                        history_values=history_values,
                    )
                    entries.append(
                        _entry(
                            parity_id=f"formula:{contract.formula_id}:{period}",
                            domain="formula_derived_metrics",
                            metric=contract.description,
                            period=str(period),
                            dimensions={},
                            legacy_range=legacy_range if contract.sheet in legacy.sheetnames else "",
                            source_kind="derived",
                            normalized_path="",
                            requirement="may_improve_semantically",
                            minimum=1,
                            writes=[],
                            formula_cell=f"{contract.sheet}!{coordinate}",
                            formula_present=formula_present,
                            formula_protected=formula_protected,
                            economically_calculable=calculable,
                            calculation_reason=calculation_reason,
                            inventory_class="formula_improvement",
                            inventory_origin="generic_formula_contract",
                            comparison_result=(
                                "formula_present_protected"
                                if formula_present and formula_protected
                                else "formula_present_unprotected"
                                if formula_present
                                else "formula_missing"
                            ),
                        )
                    )
        support = wb["{ticker}_Investment_Case_Data"]
        for _scenario_label, scenario_token, _scenario_column in CANONICAL_SCENARIOS:
            for method_id, _name_token, metric_id, _offset in CANONICAL_VALUATION_METHODS:
                row = canonical_valuation_matrix_row(scenario_token, method_id)
                formula_cells = tuple(support.cell(row, column) for column in range(57, 68))
                formula_present = all(
                    isinstance(cell.value, str) and cell.value.startswith("=")
                    for cell in formula_cells
                )
                formula_protected = all(bool(cell.protection.locked) for cell in formula_cells)
                entries.append(
                    _entry(
                        parity_id=f"investment_case_valuation:{scenario_token}:{method_id}",
                        domain="investment_case_valuation_outputs",
                        metric=metric_id,
                        period="scenario",
                        dimensions={"scenario": scenario_token, "method": method_id},
                        legacy_range="",
                        source_kind="derived",
                        normalized_path="valuation_inputs",
                        requirement="may_improve_semantically",
                        minimum=1,
                        writes=[],
                        formula_cell=f"{{ticker}}_Investment_Case_Data!BE{row}:BO{row}",
                        formula_present=formula_present,
                        formula_protected=formula_protected,
                        economically_calculable=formula_present,
                        calculation_reason=(
                            "Canonical Investment Case method formulas own all numeric and domain gates; "
                            "the row remains blank when typed inputs are unavailable."
                        ),
                        inventory_class="formula_improvement",
                        inventory_origin="investment_case_canonical_formula_contract",
                        comparison_result=(
                            "formula_present_protected"
                            if formula_present and formula_protected
                            else "formula_present_unprotected"
                            if formula_present
                            else "formula_missing"
                        ),
                    )
                )

        summary = wb["Valuation"]
        for row in range(194, 199):
            for column in range(2, 6):
                cell = summary.cell(row, column)
                coordinate = cell.coordinate
                formula_present = isinstance(cell.value, str) and cell.value.startswith("=IC_")
                formula_protected = bool(cell.protection.locked)
                entries.append(
                    _entry(
                        parity_id=f"valuation_forward_summary:{coordinate}",
                        domain="valuation_forward_summary",
                        metric=str(summary.cell(row, 1).value or coordinate),
                        period="scenario",
                        dimensions={"scenario": str(summary.cell(193, column).value or "")},
                        legacy_range="",
                        source_kind="derived_reference",
                        normalized_path="",
                        requirement="may_improve_semantically",
                        minimum=1,
                        writes=[],
                        formula_cell=f"Valuation!{coordinate}",
                        formula_present=formula_present,
                        formula_protected=formula_protected,
                        economically_calculable=formula_present,
                        calculation_reason="Direct named reference to the canonical Investment Case output.",
                        inventory_class="formula_improvement",
                        inventory_origin="canonical_summary_reference",
                        comparison_result=(
                            "formula_present_protected"
                            if formula_present and formula_protected
                            else "formula_present_unprotected"
                            if formula_present
                            else "formula_missing"
                        ),
                    )
                )
    finally:
        wb.close()
        legacy.close()

    domain_counts: dict[str, Counter[str]] = defaultdict(Counter)
    for row in entries:
        domain_counts[str(row["domain"])][str(row["current_status"])] += 1
    reproduced_statuses = {"reproduced_correctly", "reproduced_with_improved_wording"}
    required = [row for row in entries if row["parity_requirement"] == "must_reproduce"]
    missing_required = [row for row in required if row["current_status"] not in reproduced_statuses]
    independent_source_facts = [
        row for row in entries
        if row["inventory_class"] == "source_fact" and row["inventory_origin"] == "legacy_workbook_business_key"
    ]
    inventory_class_counts = Counter(str(row["inventory_class"]) for row in entries)
    formula_entries = [row for row in entries if row["inventory_class"] == "formula_improvement"]
    formula_contract_counts = Counter(str(row["formula_contract_status"]) for row in formula_entries)
    formula_calculability_counts = Counter(str(row["economic_calculability"]) for row in formula_entries)
    promise_entries = [row for row in entries if row["domain"] == "promise_progress"]
    promise_key_disposition_counts = Counter(
        str((row.get("dimensions") or {}).get("promise_parity_category") or "unavailable_without_adequate_evidence")
        for row in promise_entries
    )
    promise_occurrence_disposition_counts = Counter(
        str(row.get("disposition") or "")
        for row in (_path_get(package, "promise_progress.historical_evidence_items") or [])
        if isinstance(row, Mapping)
    )
    return {
        "$schema": "./anf_new_ticker_parity_matrix.schema.json",
        "version": "1.4.0",
        "contract_name": "ANF legacy-oracle parity matrix",
        "architectural_scope": "legacy_adapter_fixture_only; generic engine remains ticker-neutral",
        "inventory_method": "legacy workbook business keys are inventoried first; normalized package and plan are comparison targets only",
        "formula_contract_version": FORMULA_CONTRACT_VERSION,
        "source_digests": {
            "legacy_workbook_sha256": _sha(legacy_path),
            "normalized_package_sha256": hashlib.sha256(json.dumps(package, sort_keys=True, default=str).encode()).hexdigest(),
            "binding_plan_sha256": hashlib.sha256(json.dumps(plan, sort_keys=True, default=str).encode()).hexdigest(),
            "shell_sha256": _sha(shell_path),
            "binding_map_sha256": _portable_file_sha256(binding_path),
        },
        "summary": {
            "entry_count": len(entries),
            "required_count": len(required),
            "required_reproduced_count": len(required) - len(missing_required),
            "required_missing_count": len(missing_required),
            "independent_source_fact_count": len(independent_source_facts),
            "independent_source_fact_reproduced_count": sum(
                row["current_status"] in reproduced_statuses for row in independent_source_facts
            ),
            "inventory_class_counts": dict(sorted(inventory_class_counts.items())),
            "formula_contract_counts": dict(sorted(formula_contract_counts.items())),
            "formula_calculability_counts": dict(sorted(formula_calculability_counts.items())),
            "guidance_destination_lineage": guidance_destination_lineage,
            "promise_progress_key_disposition_counts": dict(sorted(promise_key_disposition_counts.items())),
            "promise_progress_occurrence_disposition_counts": dict(sorted(promise_occurrence_disposition_counts.items())),
            "domain_status_counts": {key: dict(value) for key, value in sorted(domain_counts.items())},
        },
        "entries": entries,
    }


def _markdown(matrix: Mapping[str, Any]) -> str:
    summary = matrix["summary"]
    lines = [
        "# ANF New-Ticker Parity Matrix",
        "",
        "ANF is a read-only migration oracle. This matrix locks business-key coverage without moving ANF logic into the generic engine.",
        "",
        f"- Entries: {summary['entry_count']}",
        f"- Required reproduced: {summary['required_reproduced_count']} / {summary['required_count']}",
        f"- Required missing: {summary['required_missing_count']}",
        f"- Independently inventoried legacy source facts reproduced: {summary['independent_source_fact_reproduced_count']} / {summary['independent_source_fact_count']}",
        f"- Inventory classes: {json.dumps(summary['inventory_class_counts'], sort_keys=True)}",
        f"- Formula contracts: {json.dumps(summary['formula_contract_counts'], sort_keys=True)}",
        f"- Formula calculability: {json.dumps(summary['formula_calculability_counts'], sort_keys=True)}",
        f"- Promise Progress key dispositions: {json.dumps(summary['promise_progress_key_disposition_counts'], sort_keys=True)}",
        f"- Promise Progress occurrence dispositions: {json.dumps(summary['promise_progress_occurrence_disposition_counts'], sort_keys=True)}",
        "",
        "| Domain | Economically reproduced | Contract present, blank by missing evidence | Missing / unavailable |",
        "|---|---:|---:|---:|",
    ]
    for domain, counts in summary["domain_status_counts"].items():
        lines.append(
            f"| {domain} | {counts.get('reproduced_correctly', 0)} | "
            f"{counts.get('contract_present_blank_by_missing_evidence', 0)} | "
            f"{counts.get('missing_or_explicitly_unavailable', 0)} |"
        )
    return "\n".join(lines) + "\n"


def main(argv: Sequence[str] | None = None) -> int:
    data_root = _default_data_root()
    default_dir = data_root / "outputs" / "stress_tests" / "ANF_new_ticker_engine"
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--package", type=Path, default=default_dir / "ANF_normalized_data_package.json")
    parser.add_argument("--plan", type=Path, default=default_dir / "ANF_binding_plan.json")
    parser.add_argument("--legacy", type=Path, default=data_root / "outputs" / "Excel stock models" / "ANF_model.xlsx")
    parser.add_argument("--shell", type=Path, default=DEFAULT_SHELL)
    parser.add_argument("--binding-map", type=Path, default=DEFAULT_BINDINGS)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON)
    parser.add_argument("--output-md", type=Path, default=DEFAULT_OUTPUT_MD)
    args = parser.parse_args(argv)
    matrix = build_parity_matrix(
        package=load_json_strict(args.package),
        plan=load_json_strict(args.plan),
        legacy_path=args.legacy,
        shell_path=args.shell,
        binding_path=args.binding_map,
    )
    args.output_json.write_text(json.dumps(matrix, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    args.output_md.write_text(_markdown(matrix), encoding="utf-8")
    print(f"parity matrix: {args.output_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
