"""Build the read-only ANF legacy-adapter normalized-data shadow package.

This is a migration fixture, not the generic source-native package path for new
tickers. It reads saved ANF workbook/support-sheet artifacts and source folders,
then emits normalized JSON and coverage reports. It does not call production
workbook writers and it never creates an ANF workbook.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from pbi_xbrl.normalized_company_data_validation import (  # noqa: E402
    build_mapping_gap_report,
    build_normalized_text_quality_audit,
    classify_normalized_text_quality,
    validate_normalized_company_data,
)
from pbi_xbrl.json_schema_validation import load_json_strict  # noqa: E402
from pbi_xbrl.new_ticker_guidance_scope import (  # noqa: E402
    guidance_scope_key,
    normalize_guidance_scope,
)
from pbi_xbrl.new_ticker_binding_planner import (  # noqa: E402
    DEFAULT_MANIFEST,
    DEFAULT_SHELL,
    BindingPlan,
    BindingPlanSnapshot,
    inspect_binding_eligibility,
    reproduce_binding_plan,
)
from pbi_xbrl.valuation_scenario_economics import canonicalize_scenario_contract  # noqa: E402
from pbi_xbrl.segment_normalization import (  # noqa: E402
    SegmentNormalizationError,
    SegmentSourceFact,
    canonicalize_segment_source_facts,
    canonical_segment_business_identity,
    canonical_segment_period_type,
    normalize_segment_currency_to_millions,
    segment_aggregation_role,
)


REQUIRED_SECTIONS = [
    "ticker_metadata",
    "company_profile",
    "quarterly_financials",
    "calculation_history",
    "annual_financials",
    "debt_liquidity",
    "capital_returns",
    "normalized_guidance",
    "promise_progress",
    "segments",
    "operating_drivers",
    "quarter_notes",
    "investment_case",
    "valuation_outputs",
    "source_coverage",
    "mapping_gaps",
    "manual_review_flags",
]

# Guardrail for callers and tests: generic ticker onboarding must start from
# source evidence candidates, never from a legacy workbook adapter.
LEGACY_WORKBOOK_ADAPTER_FIXTURE = True
GENERIC_SOURCE_NATIVE_BUILDER = False

SOURCE_FAMILIES = {
    "annual_reports": "annual reports",
    "earnings_release": "earnings releases",
    "earnings_presentation": "presentations",
    "earnings_transcripts": "transcripts",
    "press_release": "press releases",
    "financial_statement": "financial statement files",
    "conferences": "conference files",
}

_NORMALIZED_UNITS = {
    "$", "$m", "$bn", "USD", "USDm", "USDbn", "%", "bps", "pp", "x",
    "$/share", "m shares", "shares", "count", "days", "quarters", "pts",
    "ratio", "stores", "visits", "m visits", "units",
}

_LEGACY_PROMISE_AUDIT_ONLY_KEYS = {
    ("capital_expenditures", 2020),
    ("capital_expenditures", 2022),
    ("capital_expenditures", 2023),
}

_LEGACY_PROMISE_REJECTION_REASONS = {
    ("revenue", 2019): (
        "The cited 1% is the fiscal-2018 calendar/foreign-currency impact on reported sales, "
        "not fiscal-2019 revenue guidance."
    ),
    ("tariffs", 2019): (
        "The cited basis points describe gross-margin movement or a combined foreign-currency and tariff effect, "
        "not a definition-compatible standalone tariff promise."
    ),
    ("revenue", 2020): (
        "The cited 80% is reopened-store sales productivity versus the prior year, not fiscal-2020 revenue guidance."
    ),
    ("tariffs", 2020): (
        "The cited source is a fiscal-2019 outlook and describes a combined foreign-currency and tariff effect; "
        "it does not support a fiscal-2020 tariff promise."
    ),
    ("revenue", 2022): (
        "The cited percentages describe inventory attributes or operating-margin outlook, not fiscal-2022 revenue guidance."
    ),
}

_ANF_FY2025_PRE_RELEASE_FULL_YEAR_ROWS = {
    185: ("Revenue", "at least 6%"),
    186: ("Operating margin", "around 13%"),
    190: ("Diluted shares", "around 48 million"),
    191: ("Capex", "~ $245 million"),
    192: ("Real estate activity", "~40 net store openings"),
}
_ANF_FY2025_PRE_RELEASE_DOCUMENT = "ANF_2026-01-12_press_release_business_update.pdf"
_ANF_FY2025_PRE_RELEASE_CONTEXT = "Full Year Fiscal 2025 Outlook"

_LEGACY_REPORT_PLACEHOLDER_ROWS = {
    "total_debt": ("REPORT_BS_Q", "Total debt"),
    "debt_core": ("REPORT_BS_Q", "Debt core"),
    "interest_paid": ("REPORT_CF_Q", "Cash interest"),
}

_ANF_FY2018_BALANCE_SHEET_SOURCE_REF = (
    "tickers/ANF/earnings_release/8-K_2019-03-07_earnings_release.htm"
    "#consolidated-balance-sheets-in-thousands"
)
_ANF_HISTORY_SCALE_CORRECTIONS = {
    ("2018-Q4", "cash", 723_135_000_000): 723_135_000,
    ("2018-Q4", "inventory", 437_879_000_000): 437_879_000,
}

FINANCIAL_FIELD_DEFINITIONS = {
    "revenue": "Revenue reported for the fiscal period.",
    "cost_of_goods_sold": "Reported cost of goods sold for the fiscal period.",
    "gross_profit": "Reported revenue less reported cost of sales.",
    "operating_income": "Reported operating income for the fiscal period.",
    "base_ebitda": "EBITDA before company-defined non-GAAP adjustments.",
    "adjusted_ebitda": "Company-reported adjusted EBITDA using the source period definition.",
    "net_income": "Net income attributable to common shareholders.",
    "eps": "GAAP diluted earnings per share for the fiscal period.",
    "adjusted_eps": "Company-reported adjusted diluted earnings per share.",
    "operating_cash_flow": "Net cash provided by operating activities.",
    "capital_expenditures": "Cash capital expenditures, represented as a positive cash outflow.",
    "income_taxes_paid": "Cash income taxes paid during the fiscal period.",
    "depreciation_amortization": "Reported depreciation and amortization for the fiscal period.",
    "operating_margin": "Operating income divided by reported revenue.",
    "free_cash_flow": "Operating cash flow less cash capital expenditures.",
    "diluted_shares": "Quarterly weighted-average diluted shares used for diluted EPS.",
    "shares_outstanding": "Point-in-time common shares outstanding at the period end.",
    "total_equity": "Total shareholders' equity at the period end.",
    "book_value_per_share": "Total shareholders' equity divided by point-in-time shares outstanding.",
    "tangible_book_value_per_share": "Equity less goodwill and intangible assets divided by point-in-time shares outstanding.",
    "cash": "Cash and cash equivalents at the period end.",
    "marketable_securities": "Marketable securities reported as a separate point-in-time balance.",
    "total_assets": "Total assets reported at the period end.",
    "total_liabilities": "Total liabilities reported at the period end.",
    "current_assets": "Current assets reported at the period end.",
    "current_liabilities": "Current liabilities reported at the period end.",
    "inventory": "Inventory reported at the period end.",
    "accounts_payable": "Accounts payable reported at the period end.",
    "accrued_liabilities": "Accrued liabilities reported at the period end.",
    "property_plant_equipment_net": "Property and equipment, net, reported at the period end.",
    "other_assets_noncurrent": "Other non-current assets reported at the period end.",
    "other_liabilities_noncurrent": "Other non-current liabilities reported at the period end.",
    "lease_liabilities": "Current and non-current operating lease liabilities at the period end.",
    "lease_liabilities_current": "Current operating lease liabilities at the period end.",
    "lease_liabilities_noncurrent": "Non-current operating lease liabilities at the period end.",
    "total_debt": "Interest-bearing borrowings plus lease liabilities when the source definition includes both.",
    "debt_core": "Interest-bearing core borrowings excluding operating lease liabilities.",
    "net_debt": "Source-backed core borrowings less cash; unavailable when debt lineage is incomplete.",
}

CALCULATION_HISTORY_METRICS = (
    "revenue",
    "gross_profit",
    "operating_income",
    "base_ebitda",
    "adjusted_ebitda",
    "net_income",
    "operating_cash_flow",
    "capital_expenditures",
    "interest_paid",
    "interest_expense",
    "buybacks_cash",
    "dividends_cash",
    "acquisitions_cash",
    "debt_repayment",
    "debt_issuance",
    "cash",
    "marketable_securities",
    "debt_core",
    "diluted_shares",
    "shares_outstanding",
    "eps",
    "adjusted_eps",
)


def _default_data_root() -> Path:
    for ancestor in [REPO_ROOT, *REPO_ROOT.parents]:
        candidate = ancestor / "StockModelData"
        if candidate.exists():
            return candidate
    return REPO_ROOT.parent / "StockModelData"


def _default_workbook_path(data_root: Path) -> Path:
    return data_root / "outputs" / "Excel stock models" / "ANF_model.xlsx"


def _default_output_dir(data_root: Path) -> Path:
    return data_root / "outputs" / "stress_tests" / "ANF_new_ticker_engine"


def _load_json(path: Path) -> dict[str, Any]:
    payload = load_json_strict(path)
    if not isinstance(payload, dict):
        raise ValueError(f"JSON contract must be an object: {path}")
    return payload


def _write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _field(
    value: Any,
    *,
    status: str = "populated",
    source_ref: str = "",
    core: bool = False,
    reason: str = "",
    unit: str = "",
    period: str = "",
    confidence: str = "legacy_artifact_backed",
    definition: str = "",
    evidence_refs: Sequence[str] = (),
    evidence_classification: str = "",
    review_state: str = "",
) -> dict[str, Any]:
    out: dict[str, Any] = {
        "value": value,
        "status": status,
        "source_ref": source_ref,
        "core": bool(core),
    }
    if reason:
        out["reason"] = reason
    if unit:
        normalized_unit = unit if unit in _NORMALIZED_UNITS else "units"
        out["unit"] = normalized_unit
        if normalized_unit != unit:
            out["legacy_unit"] = unit
            out["unit_normalization_status"] = "manual_review_required"
    if period:
        out["period"] = period
    if confidence:
        out["confidence"] = confidence
    if definition:
        out["definition"] = definition
    normalized_refs = list(dict.fromkeys(str(ref).strip() for ref in evidence_refs if str(ref).strip()))
    if normalized_refs:
        out["evidence_refs"] = normalized_refs
    if evidence_classification:
        out["evidence_classification"] = evidence_classification
    if review_state:
        out["review_state"] = review_state
    return out


def _missing(reason: str, *, source_ref: str = "", core: bool = False) -> dict[str, Any]:
    return _field(None, status="missing_source", source_ref=source_ref, core=core, reason=reason, confidence="")


def _not_applicable(reason: str, *, source_ref: str = "", core: bool = False) -> dict[str, Any]:
    return _field(None, status="not_applicable", source_ref=source_ref, core=core, reason=reason, confidence="")


def _is_present(value: Any) -> bool:
    return value not in (None, "")


def _to_iso(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "")


def _normalize_period(value: Any, *, period_type: str = "quarterly") -> str:
    raw = _to_iso(value).strip()
    if re.fullmatch(r"\d{4}-(?:Q[1-4]|FY)", raw):
        return raw
    if re.fullmatch(r"FY\d{4}", raw):
        return f"{raw[2:]}-FY"
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", raw)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if period_type == "annual":
            return f"{year}-FY"
        quarter = (month - 1) // 3 + 1
        return f"{year}-Q{quarter}"
    return raw


def _publication_date_from_source(source: Any, fallback: Any = None) -> str:
    text = str(source or "")
    matches = re.findall(r"(?:19|20)\d{2}-\d{2}-\d{2}", text)
    if matches:
        return matches[-1]
    return _to_iso(fallback)


def _legacy_guidance_reporting_period(label: str, *, source_date: str, horizon: str) -> str:
    normalized_label = _normalize_period(label, period_type="quarterly")
    if re.fullmatch(r"\d{4}-Q[1-4]", normalized_label):
        return normalized_label
    horizon_year = re.search(r"(?:FY)?(20\d{2})", horizon)
    source_match = re.fullmatch(r"(20\d{2})-(\d{2})-\d{2}", source_date)
    if (
        "pre-release" in label.casefold()
        and horizon_year
        and source_match
        and int(horizon_year.group(1)) == int(source_match.group(1)) - 1
        and int(source_match.group(2)) <= 2
    ):
        return f"{horizon_year.group(1)}-Q4"
    return _normalize_period(source_date, period_type="quarterly")


def _percentage_value(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value or "").strip().replace("\u00a0", " ").replace(",", ".")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def _build_revenue_stream_rows(workbook_path: Path, *, period: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for display_order, row_number in enumerate(range(9, 12), start=1):
        member = str(_read_cell(workbook_path, "SUMMARY", f"A{row_number}") or "").strip().rstrip(":")
        raw_mix = _read_cell(workbook_path, "SUMMARY", f"B{row_number}")
        mix = _percentage_value(raw_mix)
        if not member or mix is None:
            continue
        member_ref = f"{workbook_path.name}!SUMMARY!A{row_number}"
        mix_ref = f"{workbook_path.name}!SUMMARY!B{row_number}"
        rows.append(
            {
                "member": _field(member, source_ref=member_ref, core=True),
                "mix": _field(mix, source_ref=mix_ref, core=True, unit="%", period=period),
                "unit": "%",
                "period": period,
                "source_ref": mix_ref,
                "display_order": display_order,
            }
        )
    return rows


def _read_legacy_valuation_series(workbook_path: Path, row_number: int) -> dict[str, float]:
    """Read an ANF legacy display row as migration evidence only."""

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb["Valuation"]
        values: dict[str, float] = {}
        for column in range(2, 14):
            period = str(ws.cell(6, column).value or "")
            value = ws.cell(row_number, column).value
            if period and isinstance(value, (int, float)):
                values[period] = float(value)
        return values
    finally:
        wb.close()


def _read_legacy_unsupported_zero_placeholders(
    workbook_path: Path,
) -> dict[tuple[str, str], dict[str, Any]]:
    """Read zero candidates that the legacy report explicitly marks missing/failed."""

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        placeholders: dict[tuple[str, str], dict[str, Any]] = {}
        for metric, (sheet_name, line_item) in _LEGACY_REPORT_PLACEHOLDER_ROWS.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            row_number = next(
                (
                    row
                    for row in range(4, ws.max_row + 1)
                    if str(ws.cell(row, 2).value or "").strip().casefold() == line_item.casefold()
                ),
                None,
            )
            if row_number is None:
                continue
            source_status = str(ws.cell(row_number, 3).value or "").strip()
            qa_status = str(ws.cell(row_number, 4).value or "").strip()
            source_missing = source_status.casefold() == "missing"
            qa_failed = qa_status.casefold() == "fail"
            if not (source_missing or qa_failed):
                continue
            for column in range(7, ws.max_column + 1):
                value = ws.cell(row_number, column).value
                if (
                    not isinstance(value, (int, float))
                    or isinstance(value, bool)
                    or float(value) != 0.0
                ):
                    continue
                period_end = _to_iso(ws.cell(3, column).value)
                if not period_end:
                    continue
                value_cell = f"{get_column_letter(column)}{row_number}"
                placeholders[(metric, period_end)] = {
                    "metric": metric,
                    "line_item": line_item,
                    "sheet": sheet_name,
                    "period_end": period_end,
                    "candidate_value": 0.0,
                    "source_status": source_status,
                    "qa_status": qa_status,
                    "value_source_ref": f"{workbook_path.name}!{sheet_name}!{value_cell}",
                    "metadata_source_ref": f"{workbook_path.name}!{sheet_name}!C{row_number}:D{row_number}",
                }
        return placeholders
    finally:
        wb.close()


def _unsupported_zero_placeholder(
    placeholders: Mapping[tuple[str, str], Mapping[str, Any]],
    metric: str,
    row: Mapping[str, Any],
    value: Any,
) -> Mapping[str, Any] | None:
    if (
        not isinstance(value, (int, float))
        or isinstance(value, bool)
        or float(value) != 0.0
    ):
        return None
    return placeholders.get((metric, _to_iso(row.get("quarter"))))


def _placeholder_source_ref(history_source_ref: str, placeholder: Mapping[str, Any]) -> str:
    return " + ".join(
        filter(
            None,
            (
                history_source_ref,
                str(placeholder.get("value_source_ref") or ""),
                str(placeholder.get("metadata_source_ref") or ""),
            ),
        )
    )


def _placeholder_reason(placeholder: Mapping[str, Any]) -> str:
    return (
        f"Legacy report line {placeholder.get('line_item')!r} marks the zero candidate "
        f"Source={placeholder.get('source_status') or 'blank'} and "
        f"QA={placeholder.get('qa_status') or 'blank'}; zero was treated as a missing placeholder."
    )


def _missing_placeholder_field(
    *,
    placeholder: Mapping[str, Any],
    history_source_ref: str,
    unit: str,
    period: str,
    core: bool = False,
) -> dict[str, Any]:
    return _field(
        None,
        status="missing_source",
        source_ref=_placeholder_source_ref(history_source_ref, placeholder),
        core=core,
        reason=_placeholder_reason(placeholder),
        unit=unit,
        period=period,
        confidence="",
    )


def _quality_checked_number(
    *,
    value: Any,
    metric: str,
    row: Mapping[str, Any],
    period: str,
    history_source_ref: str,
    placeholders: Mapping[tuple[str, str], Mapping[str, Any]],
    review_flags: list[dict[str, Any]] | None,
    core: bool = False,
) -> dict[str, Any]:
    placeholder = _unsupported_zero_placeholder(placeholders, metric, row, value)
    if placeholder is None:
        return _populated_number(value, history_source_ref, "$m", period, core=core)
    field = _missing_placeholder_field(
        placeholder=placeholder,
        history_source_ref=history_source_ref,
        unit="$m",
        period=period,
        core=core,
    )
    if review_flags is not None:
        review_flags.append(
            {
                "severity": "P2",
                "rule_id": "legacy_adapter_unsupported_zero_placeholder",
                "issue_type": "actionable_exception",
                "section": "quarterly_financials",
                "field": f"quarterly_financials.rows[{period}].{metric}",
                "normalized_path": f"quarterly_financials.rows[{period}].{metric}",
                "row_key": f"{period}|{metric}",
                "affected_period": period,
                "message": str(field["reason"]),
                "source_ref": str(field["source_ref"]),
                "root_cause": "legacy_missing_or_failed_zero_placeholder",
                "visibility_disposition": "needs_review",
                "promotion_blocking": False,
                "suggested_action": (
                    "Provide independent source-backed evidence before treating the zero candidate as a financial fact."
                ),
                "adapter_metadata": dict(placeholder),
            }
        )
    return field


def _segment_dimension(member: str) -> str:
    lower = member.strip().lower()
    if lower in {"americas", "emea", "apac", "united states", "international"}:
        return "geography"
    if lower in {"abercrombie", "hollister", "gilly hicks"}:
        return "brand"
    if lower in {"total company", "company total", "total"}:
        return "total_company"
    return "reported_segment"


def _driver_type(group: str, driver: str) -> str:
    text = f"{group} {driver}".lower()
    for token, driver_type in (
        ("demand", "demand"),
        ("price", "pricing"),
        ("volume", "volume"),
        ("margin", "margin"),
        ("cost", "cost"),
        ("capital", "capital_allocation"),
        ("liquidity", "liquidity"),
        ("strategy", "strategy"),
    ):
        if token in text:
            return driver_type
    return "operational"


def _evidence_key(*parts: Any) -> str:
    canonical = "|".join(re.sub(r"\s+", " ", str(part or "")).strip() for part in parts)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:16]


def _normalized_scalar(value: Any) -> Any:
    if isinstance(value, Mapping) and "status" in value:
        return value.get("value") if str(value.get("status") or "") == "populated" else None
    return value


def _collect_row_source_refs(value: Any) -> list[str]:
    refs: set[str] = set()

    def visit(node: Any) -> None:
        if isinstance(node, Mapping):
            direct = node.get("source_ref")
            if isinstance(direct, str) and direct.strip():
                refs.add(direct.strip())
            source = node.get("source")
            if isinstance(source, str) and source.strip():
                refs.add(source.strip())
            elif isinstance(source, Mapping):
                for key in ("doc", "path", "file", "url", "source_ref"):
                    candidate = source.get(key)
                    if isinstance(candidate, str) and candidate.strip():
                        refs.add(candidate.strip())
            for child in node.values():
                visit(child)
        elif isinstance(node, list):
            for child in node:
                visit(child)

    visit(value)
    return sorted(refs)


def _legacy_adapter_business_key(row: Mapping[str, Any], collection_path: str, source_index: int) -> str:
    field_sets = {
        "quarterly_financials.rows": ("period",),
        "annual_financials.rows": ("period",),
        "normalized_guidance.items": ("metric", "horizon", "publication_date", "evidence_key"),
        "segments.items": ("period", "dimension", "member", "metric", "evidence_key"),
        "operating_drivers.items": ("period", "topic", "driver_type", "evidence_key"),
        "quarter_notes.items": ("period", "theme", "metric", "evidence_key"),
    }
    parts = [
        str(value).strip()
        for field in field_sets.get(collection_path, ("period", "evidence_key"))
        if (value := _normalized_scalar(row.get(field))) not in (None, "")
    ]
    return "|".join(parts) or f"source_index:{source_index}"


def _legacy_adapter_truncation_detail(
    row: Mapping[str, Any],
    *,
    collection_path: str,
    source_index: int,
    truncation_index: int,
    detail_index: int,
    limit: int,
    fallback_source_ref: str,
) -> dict[str, Any]:
    source_refs = _collect_row_source_refs(row)
    period = next(
        (
            str(value)
            for field in ("period", "quarter", "horizon", "stated_in_period", "publication_date")
            if (value := _normalized_scalar(row.get(field))) not in (None, "")
        ),
        "",
    )
    evidence_key = str(_normalized_scalar(row.get("evidence_key")) or "")
    business_row_key = _legacy_adapter_business_key(row, collection_path, source_index)
    lineage_material = "|".join(
        [
            collection_path,
            business_row_key,
            period,
            evidence_key,
            *source_refs,
            str(source_index),
        ]
    )
    return {
        "collection": collection_path,
        "section": collection_path.split(".", 1)[0],
        "detail_path": (
            f"source_coverage.legacy_adapter_truncations.{truncation_index}."
            f"excluded_rows.{detail_index}"
        ),
        "adapter_candidate_path": f"legacy_adapter_candidates.{collection_path}.{source_index}",
        "lineage_id": f"TRUNC-{hashlib.sha256(lineage_material.encode('utf-8')).hexdigest()[:24]}",
        "source_index": source_index,
        "business_row_key": business_row_key,
        "period": period,
        "evidence_key": evidence_key,
        "source_ref": source_refs[0] if source_refs else fallback_source_ref,
        "source_refs": source_refs,
        "truncation_rule": f"tail:{limit}",
        "reason": "Row falls outside the explicit legacy-adapter tail limit and remains audit-only.",
    }


def _limit_legacy_adapter_rows(
    rows: Sequence[Mapping[str, Any]],
    *,
    limit: int,
    collection_path: str,
    workbook_path: Path,
    truncations: list[dict[str, Any]],
    review_flags: list[dict[str, Any]],
) -> list[Mapping[str, Any]]:
    materialized = list(rows)
    if len(materialized) <= limit:
        return materialized
    dropped = len(materialized) - limit
    source_ref = f"{workbook_path.name}!legacy_adapter_selection"
    truncation_index = len(truncations)
    excluded_rows = [
        _legacy_adapter_truncation_detail(
            row,
            collection_path=collection_path,
            source_index=detail_index,
            truncation_index=truncation_index,
            detail_index=detail_index,
            limit=limit,
            fallback_source_ref=source_ref,
        )
        for detail_index, row in enumerate(materialized[:-limit])
    ]
    record = {
        "collection": collection_path,
        "selection": "tail",
        "input_rows": len(materialized),
        "retained_rows": limit,
        "dropped_rows": dropped,
        "excluded_row_count": len(excluded_rows),
        "excluded_rows": excluded_rows,
        "truncation_rule": f"tail:{limit}",
        "reason": "Legacy migration fixture capacity limit; source-native builders must not copy this policy.",
        "source_ref": source_ref,
    }
    detail_ref = f"source_coverage.legacy_adapter_truncations.{len(truncations)}.excluded_rows"
    truncations.append(record)
    review_flags.append(
        {
            "severity": "P2",
            "rule_id": "legacy_adapter_truncation",
            "field": collection_path,
            "message": f"Legacy adapter retained the latest {limit} of {len(materialized)} rows and explicitly recorded {dropped} dropped rows.",
            "source_ref": source_ref,
            "suggested_action": "Use source-native evidence selection and an explicit planner overflow policy before onboarding a new ticker.",
            "adapter_metadata": {
                "collection": collection_path,
                "input_rows": len(materialized),
                "retained_rows": limit,
                "dropped_rows": dropped,
                "truncation_rule": f"tail:{limit}",
                "detail_ref": detail_ref,
            },
        }
    )
    return materialized[-limit:]


def _dedupe_legacy_adapter_rows(
    rows: Sequence[Mapping[str, Any]],
    *,
    collection_path: str,
    workbook_path: Path,
    deduplications: list[dict[str, Any]],
    review_flags: list[dict[str, Any]],
) -> list[Mapping[str, Any]]:
    retained: list[Mapping[str, Any]] = []
    seen: set[str] = set()
    for source_index, row in enumerate(rows):
        evidence_key = str(row.get("evidence_key") or "")
        if not evidence_key or evidence_key not in seen:
            retained.append(row)
            if evidence_key:
                seen.add(evidence_key)
            continue
        source_ref = f"{workbook_path.name}!legacy_adapter_selection"
        record = {
            "collection": collection_path,
            "source_index": source_index,
            "evidence_key": evidence_key,
            "reason": "Exact duplicate evidence key in the legacy workbook projection.",
            "source_ref": source_ref,
        }
        deduplications.append(record)
        review_flags.append(
            {
                "severity": "P2",
                "rule_id": "legacy_adapter_exact_duplicate",
                "field": f"{collection_path}.{source_index}",
                "message": f"Legacy adapter omitted an exact duplicate evidence row with key {evidence_key}.",
                "source_ref": source_ref,
                "suggested_action": "Confirm source-native evidence identity and deduplication policy before promotion.",
                "adapter_metadata": record,
            }
        )
    return retained


def _to_millions(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if abs(numeric) >= 100_000:
        return round(numeric / 1_000_000, 3)
    return round(numeric, 3)


def _anf_history_source_evidence(
    row: Mapping[str, Any],
    source_field: str,
    legacy_source_ref: str,
) -> tuple[Any, str]:
    """Correct two source-proven FY2018 scale defects in the ANF fixture."""

    raw_value = row.get(source_field)
    if not isinstance(raw_value, (int, float)) or isinstance(raw_value, bool):
        return raw_value, legacy_source_ref
    key = (str(row.get("fiscal_label") or ""), source_field, raw_value)
    corrected_value = _ANF_HISTORY_SCALE_CORRECTIONS.get(key)
    if corrected_value is None:
        return raw_value, legacy_source_ref
    return corrected_value, f"{legacy_source_ref} + {_ANF_FY2018_BALANCE_SHEET_SOURCE_REF}"


def _clean_text(value: Any, *, limit: int = 420) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "…"


def _text_quality_action(classification: str) -> str:
    actions = {
        "boilerplate_or_legal": "Keep as source/audit coverage only; replace visible text with source-backed quarter commentary.",
        "compensation_or_governance_noise": "Keep governance/compensation snippets out of visible quarter notes.",
        "accounting_policy_or_definition": "Do not use formula/accounting definitions as operating-driver reads.",
        "release_header_or_source_title": "Strip source headers from visible segment notes.",
        "fragmented_sentence": "Review parser boundaries and rebuild a complete source-backed sentence.",
        "too_long_unstructured": "Condense or demote the snippet before visible rendering.",
        "missing_context": "Keep out of visible UI until enough context is normalized.",
    }
    return actions.get(classification, "Review this text before visible rendering.")


def _record_text_quality_demotion(
    demotions: list[dict[str, Any]],
    *,
    section: str,
    field: str,
    source_ref: str,
    classification: str,
    text: str,
) -> None:
    demotions.append(
        {
            "severity": "P2",
            "rule_id": "text_quality_demoted",
            "field": field,
            "message": f"Demoted non-visible-ready {section} text: {classification}.",
            "source_ref": source_ref,
            "suggested_action": _text_quality_action(classification),
            "section": section,
            "classification": classification,
            "original_excerpt": _clean_text(text, limit=220),
        }
    )


def _visible_text_or_blank(
    text: str,
    *,
    field: str,
    section: str,
    source_ref: str,
    demotions: list[dict[str, Any]],
) -> str:
    if not text:
        return ""
    classification = classify_normalized_text_quality(text, field=field, visible_ui=True)
    if classification == "clean_visible_ui":
        return text
    _record_text_quality_demotion(
        demotions,
        section=section,
        field=field,
        source_ref=source_ref,
        classification=classification,
        text=text,
    )
    return ""


def _visible_text_is_clean(
    text: str,
    *,
    field: str,
    section: str,
    source_ref: str,
    demotions: list[dict[str, Any]],
) -> bool:
    if not text:
        _record_text_quality_demotion(
            demotions,
            section=section,
            field=field,
            source_ref=source_ref,
            classification="missing_context",
            text="",
        )
        return False
    classification = classify_normalized_text_quality(text, field=field, visible_ui=True)
    if classification == "clean_visible_ui":
        return True
    _record_text_quality_demotion(
        demotions,
        section=section,
        field=field,
        source_ref=source_ref,
        classification=classification,
        text=text,
    )
    return False


def _text_quality_demotion_summary(demotions: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    by_section: dict[str, int] = defaultdict(int)
    by_classification: dict[str, int] = defaultdict(int)
    for row in demotions:
        by_section[str(row.get("section") or "unknown")] += 1
        by_classification[str(row.get("classification") or "unknown")] += 1
    return {
        "total_demoted": len(demotions),
        "by_section": dict(sorted(by_section.items())),
        "by_classification": dict(sorted(by_classification.items())),
    }


def _read_sheet_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        headers = [str(value or "").strip() for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows: list[dict[str, Any]] = []
        for row_idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = {headers[col_idx]: value for col_idx, value in enumerate(values) if col_idx < len(headers) and headers[col_idx]}
            if any(_is_present(value) for value in row.values()):
                row["_row_number"] = row_idx
                rows.append(row)
        return rows
    finally:
        wb.close()


def _read_cell(workbook_path: Path, sheet_name: str, cell_ref: str) -> Any:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return wb[sheet_name][cell_ref].value
    finally:
        wb.close()


def _anf_transcript_ref(start_line: int, end_line: int | None = None) -> str:
    suffix = f"L{start_line}" if end_line is None or end_line == start_line else f"L{start_line}-L{end_line}"
    return f"tickers/ANF/earnings_transcripts/ANF_Q4_2025_transcript.txt#{suffix}"


def _anf_annual_report_ref(page: int) -> str:
    return f"tickers/ANF/annual_reports/ANF_2025_annual_report.pdf#page={page}"


def _narrative_field(
    value: str,
    evidence_refs: Sequence[str],
    *,
    classification: str,
    core: bool = False,
    review_state: str = "accepted",
) -> dict[str, Any]:
    refs = list(dict.fromkeys(ref for ref in evidence_refs if ref))
    return _field(
        value,
        source_ref=refs[0] if refs else "",
        core=core,
        evidence_refs=refs,
        evidence_classification=classification,
        review_state=review_state,
    )


def _build_anf_source_backed_operating_drivers(period: str) -> dict[str, Any]:
    specs = (
        (
            "Sales execution",
            "Management guides to 3%-5% sales growth in 2026 and says product execution, marketing and store experience determine the range of outcomes.",
            "The growth case depends on both brands converting traffic into demand while lapping stronger prior-year comparisons.",
            (_anf_transcript_ref(52, 52), _anf_transcript_ref(313, 319)),
            "demand",
        ),
        (
            "Margin durability",
            "The 2026 operating-margin guide is 12.0%-12.5%; tariffs, ERP and marketing are partly offset by freight and modest AUR improvement.",
            "Margin delivery is the central test of whether recent profitability is durable rather than peak-cycle.",
            (_anf_transcript_ref(54, 60), _anf_transcript_ref(290, 292)),
            "margin",
        ),
        (
            "Inventory quality",
            "Year-end inventory units were up 5%, including about 3 points of ERP prebuild; management said underlying units were up about 2% and both brands were in chase position.",
            "A clean chase position lowers markdown risk, while ERP timing and tariff costs still need monitoring.",
            (_anf_transcript_ref(42, 42), _anf_transcript_ref(303, 305)),
            "operational",
        ),
        (
            "Capital returns",
            "Fiscal 2025 free cash flow was $378 million versus $450 million of repurchases, and management targets about $450 million of repurchases in 2026.",
            "Repurchases support per-share results, but spending above free cash flow makes liquidity discipline an explicit watch item.",
            (_anf_transcript_ref(48, 48), _anf_transcript_ref(56, 58)),
            "capital_allocation",
        ),
    )
    items: list[dict[str, Any]] = []
    for priority, (topic, current_read, why, refs, driver_type) in enumerate(specs, start=1):
        items.append(
            {
                "topic": _narrative_field(topic, refs, classification="source_backed_fact", core=True),
                "driver": _narrative_field(topic, refs, classification="source_backed_fact", core=True),
                "current_read": _narrative_field(current_read, refs, classification="evidence_backed_synthesis", core=True),
                "metric_value": _missing("The visible watchlist is a qualitative evidence synthesis.", source_ref=refs[0]),
                "source": refs[0],
                "why_it_matters": _narrative_field(why, refs, classification="evidence_backed_synthesis", core=True),
                "quality": "source_backed_curated_narrative",
                "period": period,
                "horizon": "2026-FY",
                "driver_type": driver_type,
                "evidence_key": _evidence_key("anf_driver", topic, *refs),
                "evidence_refs": list(refs),
                "review_state": "accepted",
                "display_role": "current_watchlist",
                "display_priority": priority,
            }
        )

    current_outlook = {
        "current_actual_read": _narrative_field(
            "Q4 sales grew 5%, operating margin was 14.1% and diluted EPS was $3.68.",
            (_anf_transcript_ref(14, 16),),
            classification="source_backed_fact",
        ),
        "current_actual_use": _narrative_field(
            "Use Q4 2025 as the latest reported baseline for sales, margin and earnings momentum.",
            (_anf_transcript_ref(14, 16),),
            classification="evidence_backed_synthesis",
        ),
        "current_guidance_read": _narrative_field(
            "For 2026, management guides to 3%-5% sales growth, 12.0%-12.5% operating margin and adjusted EPS of $10.20-$11.00.",
            (_anf_transcript_ref(52, 56),),
            classification="source_backed_fact",
        ),
        "current_guidance_use": _narrative_field(
            "Track brand growth and margin delivery against the full-year ranges without mixing them with Q1 guidance.",
            (_anf_transcript_ref(52, 62),),
            classification="evidence_backed_synthesis",
        ),
        "margin_bridge_read": _narrative_field(
            "Q1 includes about 290 bps of tariff pressure, more than 100 bps of ERP impact and 50 bps of marketing, partly offset by roughly 160 bps of freight benefit and modest AUR improvement.",
            (_anf_transcript_ref(58, 62), _anf_transcript_ref(290, 292)),
            classification="source_backed_fact",
        ),
        "margin_bridge_use": _narrative_field(
            "The quarter tests whether freight, pricing and sourcing mitigation can offset temporary and structural cost pressure.",
            (_anf_transcript_ref(54, 60), _anf_transcript_ref(290, 292)),
            classification="evidence_backed_synthesis",
        ),
    }
    return {"items": items, "current_outlook": current_outlook}


def _build_anf_source_backed_quarter_notes(period: str) -> dict[str, Any]:
    specs = (
        (
            "Q4 results",
            "Q4 sales grew 5%; operating margin was 14.1% despite 360 bps of tariff pressure, and diluted EPS reached $3.68.",
            "The quarter finished at the high end of the January update and showed balanced growth across brands and regions.",
            "Treat the quarter as a strong actual baseline, while separating reported profitability from the lower 2026 margin guide.",
            (_anf_transcript_ref(14, 16), _anf_transcript_ref(36, 42)),
        ),
        (
            "Brand mix",
            "Hollister grew 6% in Q4 and 15% for the year; Abercrombie returned to 4% Q4 growth after declining 1% for the full year.",
            "Hollister remains the growth engine, while Abercrombie's return to growth broadens the earnings setup.",
            "Watch whether Abercrombie sustains growth as Hollister laps two years of strong expansion.",
            (_anf_transcript_ref(22, 24), _anf_transcript_ref(38, 44)),
        ),
        (
            "Inventory",
            "Inventory cost and units ended 5% higher, with about 3 points tied to tariffs and the ERP prebuild; underlying units were about 2% higher.",
            "Management described both brands as being in chase position, which is more constructive than broad excess inventory.",
            "Monitor markdowns and AUR after the ERP-related inventory timing normalizes.",
            (_anf_transcript_ref(42, 42), _anf_transcript_ref(303, 305)),
        ),
        (
            "2026 margin bridge",
            "The 2026 operating-margin guide is 12.0%-12.5%; Q1 includes tariff, ERP and marketing headwinds partly offset by freight and modest AUR improvement.",
            "The bridge explains why guided margins step down even though management still expects double-digit profitability.",
            "Margin delivery is the clearest near-term proof point for valuation and earnings durability.",
            (_anf_transcript_ref(54, 62), _anf_transcript_ref(290, 292)),
        ),
        (
            "Capital allocation",
            "Fiscal 2025 free cash flow was $378 million, repurchases were $450 million and year-end liquidity was about $1.2 billion; the 2026 repurchase target is about $450 million.",
            "Buybacks support per-share results but exceeded annual free cash flow, increasing the importance of cash-generation discipline.",
            "Track free cash flow and liquidity rather than assuming unsupported debt or net-debt values.",
            (_anf_transcript_ref(48, 48), _anf_transcript_ref(56, 58)),
        ),
        (
            "Growth channels",
            "Digital represented 44% of 2025 sales, the company exceeded one billion digital visits and delivered 120 new store experiences.",
            "Stores, digital and third-party channels give the company several growth paths beyond comparable sales alone.",
            "Watch APAC capital efficiency and whether channel expansion supports the 3%-5% 2026 sales range.",
            (_anf_transcript_ref(26, 28), _anf_transcript_ref(50, 52)),
        ),
    )
    items: list[dict[str, Any]] = []
    for priority, (theme, commentary, why, implication, refs) in enumerate(specs, start=1):
        source_display = _narrative_field(
            "Q4 2025 earnings call",
            refs,
            classification="source_backed_fact",
        )
        items.append(
            {
                "theme": _narrative_field(theme, refs, classification="source_backed_fact"),
                "quarter": _field(period, source_ref=refs[0], evidence_refs=refs),
                "metric": _narrative_field(theme, refs, classification="source_backed_fact"),
                "note": _narrative_field(commentary, refs, classification="source_backed_fact", core=True),
                "commentary": _narrative_field(commentary, refs, classification="source_backed_fact", core=True),
                "why_it_matters": _narrative_field(why, refs, classification="evidence_backed_synthesis", core=True),
                "model_implication": _narrative_field(implication, refs, classification="evidence_backed_synthesis", core=True),
                "valuation_implication": _narrative_field(implication, refs, classification="evidence_backed_synthesis"),
                "source": refs[0],
                "source_display": source_display,
                "confidence": "source_backed_with_reviewed_synthesis",
                "review_state": "accepted",
                "evidence_refs": list(refs),
                "evidence_key": _evidence_key("anf_quarter_note", period, theme, *refs),
                "display_role": "current_note",
                "display_priority": priority,
            }
        )

    summary = {
        "model_read": _narrative_field(
            "ANF exits 2025 with balanced growth and strong cash generation, while 2026 depends on sustaining double-digit margins through tariff and ERP pressure.",
            (_anf_transcript_ref(14, 20), _anf_transcript_ref(52, 62)),
            classification="evidence_backed_synthesis",
        ),
        "what_changed": _narrative_field(
            "Abercrombie returned to Q4 growth, Hollister remained strong and management introduced 2026 guidance for 3%-5% sales growth and 12.0%-12.5% operating margin.",
            (_anf_transcript_ref(22, 24), _anf_transcript_ref(52, 56)),
            classification="evidence_backed_synthesis",
        ),
        "watch_next": _narrative_field(
            "Watch the Q1 tariff, ERP, freight and marketing bridge together with brand-level sales execution.",
            (_anf_transcript_ref(58, 62), _anf_transcript_ref(290, 292)),
            classification="analyst_interpretation_requiring_review",
            review_state="manual_review_required",
        ),
        "key_caveat": _narrative_field(
            "The 2026 margin guide is below 2025 adjusted performance and assumes successful mitigation of meaningful cost pressure.",
            (_anf_transcript_ref(46, 46), _anf_transcript_ref(54, 60)),
            classification="evidence_backed_synthesis",
        ),
    }
    return {"items": items, "summary": summary}


def _source_ref(sheet: str, row: Mapping[str, Any] | None = None, *, workbook_path: Path) -> str:
    if row and row.get("_row_number"):
        return f"{workbook_path.name}!{sheet}!row:{row['_row_number']}"
    return f"{workbook_path.name}!{sheet}"


def _source_file_candidates(data_root: Path) -> list[Path]:
    roots = [
        data_root / "tickers" / "ANF",
        data_root / "sec_cache" / "ANF",
        data_root / "tickers" / "sec_cache" / "ANF",
    ]
    suffixes = {".txt", ".htm", ".html", ".json", ".csv", ".xlsx", ".pdf", ".md"}
    paths: list[Path] = []
    for root in roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if path.is_file() and path.suffix.lower() in suffixes:
                paths.append(path)
    return sorted(paths, key=lambda item: str(item).lower())


def _source_coverage(data_root: Path, workbook_path: Path) -> dict[str, Any]:
    files = _source_file_candidates(data_root)
    family_counts: dict[str, int] = {}
    for family in SOURCE_FAMILIES:
        family_root = data_root / "tickers" / "ANF" / family
        family_counts[family] = len([path for path in files if family_root in path.parents])
    return {
        "sources": [
            {
                "path": str(path),
                "kind": path.suffix.lower().lstrip(".") or "file",
                "family": _classify_source_family(path),
                "status": "available",
            }
            for path in files
        ],
        "source_roots": {
            "ticker_root": str(data_root / "tickers" / "ANF"),
            "sec_cache": str(data_root / "sec_cache" / "ANF"),
            "ticker_sec_cache": str(data_root / "tickers" / "sec_cache" / "ANF"),
            "legacy_workbook": str(workbook_path),
        },
        "family_counts": family_counts,
        "legacy_artifacts": [
            "ANF_model.xlsx",
            "ANF_model.xlsx!History_Q",
            "ANF_model.xlsx!Guidance_Normalized",
            "ANF_model.xlsx!Promise_Progress",
            "ANF_model.xlsx!Quarter_Notes",
            "ANF_model.xlsx!Slides_Segments",
            "ANF_model.xlsx!Leverage_Liquidity",
            "ANF_model.xlsx!ANF_Investment_Case_Data",
        ],
    }


def _classify_source_family(path: Path) -> str:
    lower_parts = {part.lower() for part in path.parts}
    for family in SOURCE_FAMILIES:
        if family.lower() in lower_parts:
            return family
    if "sec_cache" in lower_parts:
        return "sec_xbrl"
    return "other"


def _build_quarterly_financial_rows(
    history_rows: Sequence[Mapping[str, Any]],
    workbook_path: Path,
    *,
    unsupported_zero_placeholders: Mapping[tuple[str, str], Mapping[str, Any]] | None = None,
    review_flags: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    populated = [row for row in history_rows if _is_present(row.get("revenue"))]
    populated.sort(key=lambda row: _to_iso(row.get("quarter")))
    out: list[dict[str, Any]] = []
    placeholders = unsupported_zero_placeholders or {}
    adjusted_ebitda_by_period = _read_legacy_valuation_series(workbook_path, 24)
    adjusted_eps_by_period = _read_legacy_valuation_series(workbook_path, 110)
    revolver_by_period = _read_legacy_valuation_series(workbook_path, 95)
    for row in populated:
        period = _normalize_period(row.get("fiscal_label") or row.get("quarter"), period_type="quarterly")
        source_ref = _source_ref("History_Q", row, workbook_path=workbook_path)
        cash_value, cash_source_ref = _anf_history_source_evidence(row, "cash", source_ref)
        inventory_value, inventory_source_ref = _anf_history_source_evidence(row, "inventory", source_ref)
        cfo = _to_millions(row.get("cfo"))
        capex = _to_millions(row.get("capex"))
        fcf = round(cfo - capex, 3) if cfo is not None and capex is not None else None
        eps_value = row.get("eps_diluted")
        eps_source_ref = source_ref
        if not _is_present(eps_value) and _is_present(row.get("net_income")) and _is_present(row.get("shares_diluted")):
            eps_value = float(row["net_income"]) / float(row["shares_diluted"])
            eps_source_ref = f"{source_ref} [derived: net_income / diluted_shares]"
        out.append(
            {
                "period": period,
                "fiscal_year": row.get("fiscal_year"),
                "fiscal_quarter": row.get("fiscal_quarter"),
                "period_end": _to_iso(row.get("quarter")),
                "revenue": _populated_number(row.get("revenue"), source_ref, "$m", period, core=True),
                "cost_of_goods_sold": _populated_number(row.get("cogs"), source_ref, "$m", period),
                "gross_profit": _populated_number(row.get("gross_profit"), source_ref, "$m", period),
                "operating_income": _populated_number(row.get("op_income"), source_ref, "$m", period, core=True),
                "base_ebitda": _populated_number(row.get("ebitda"), source_ref, "$m", period),
                "adjusted_ebitda": _field(
                    adjusted_ebitda_by_period[period],
                    source_ref=f"{workbook_path.name}!Valuation!row:24",
                    unit="$m",
                    period=period,
                )
                if period in adjusted_ebitda_by_period
                else _missing("Legacy adjusted EBITDA row has no value for this period.", source_ref=f"{workbook_path.name}!Valuation!row:24"),
                "net_income": _populated_number(row.get("net_income"), source_ref, "$m", period),
                "eps": _populated_scalar(eps_value, eps_source_ref, "$/share", period),
                "adjusted_eps": _field(
                    adjusted_eps_by_period[period],
                    source_ref=f"{workbook_path.name}!Valuation!row:110",
                    unit="$/share",
                    period=period,
                )
                if period in adjusted_eps_by_period
                else _missing("Legacy adjusted diluted EPS row has no value for this period.", source_ref=f"{workbook_path.name}!Valuation!row:110"),
                "operating_cash_flow": _populated_number(row.get("cfo"), source_ref, "$m", period),
                "capital_expenditures": _populated_number(row.get("capex"), source_ref, "$m", period),
                "income_taxes_paid": _populated_number(row.get("tax_paid"), source_ref, "$m", period),
                "depreciation_amortization": _populated_number(row.get("da"), source_ref, "$m", period),
                "operating_margin": _populated_scalar(row.get("operating_margin"), source_ref, "%", period),
                "free_cash_flow": _field(fcf, source_ref=source_ref, core=True, unit="$m", period=period)
                if fcf is not None
                else _missing("CFO or capex is absent for this quarter.", source_ref=source_ref, core=True),
                "diluted_shares": _populated_share_count(row.get("shares_diluted"), source_ref, period, core=True),
                "interest_paid": _quality_checked_number(
                    value=row.get("interest_paid"),
                    metric="interest_paid",
                    row=row,
                    period=period,
                    history_source_ref=source_ref,
                    placeholders=placeholders,
                    review_flags=review_flags,
                ),
                "interest_expense": _populated_number(row.get("interest_expense_net"), source_ref, "$m", period),
                "buybacks_cash": _populated_number(row.get("buybacks_cash"), source_ref, "$m", period),
                "dividends_cash": _populated_number(row.get("dividends_cash"), source_ref, "$m", period),
                "acquisitions_cash": _populated_number(row.get("acquisitions_cash"), source_ref, "$m", period),
                "debt_repayment": _populated_number(row.get("debt_repayment"), source_ref, "$m", period),
                "debt_issuance": _populated_number(row.get("debt_issuance"), source_ref, "$m", period),
                "cash": _populated_number(cash_value, cash_source_ref, "$m", period),
                "restricted_cash": _missing("Restricted cash is not separately identified in the ANF legacy history.", source_ref=source_ref),
                "short_term_investments": _populated_number(row.get("short_term_investments"), source_ref, "$m", period),
                "marketable_securities": _populated_number(row.get("marketable_securities"), source_ref, "$m", period),
                "total_assets": _populated_number(row.get("assets"), source_ref, "$m", period),
                "total_liabilities": _populated_number(row.get("liabilities"), source_ref, "$m", period),
                "current_assets": _populated_number(row.get("assets_current"), source_ref, "$m", period),
                "current_liabilities": _populated_number(row.get("liabilities_current"), source_ref, "$m", period),
                "accounts_receivable": _populated_number(row.get("accounts_receivable"), source_ref, "$m", period),
                "inventory": _populated_number(inventory_value, inventory_source_ref, "$m", period),
                "accounts_payable": _populated_number(row.get("accounts_payable_current"), source_ref, "$m", period),
                "accrued_liabilities": _populated_number(row.get("accrued_liabilities_current"), source_ref, "$m", period),
                "short_term_borrowings": _missing("Short-term borrowings are not separately identified in the ANF legacy history.", source_ref=source_ref),
                "debt_current": _populated_number(row.get("debt_current"), source_ref, "$m", period),
                "property_plant_equipment_net": _populated_number(row.get("property_plant_equipment_net"), source_ref, "$m", period),
                "other_assets_noncurrent": _populated_number(row.get("other_assets_noncurrent"), source_ref, "$m", period),
                "other_liabilities_noncurrent": _populated_number(row.get("other_liabilities_noncurrent"), source_ref, "$m", period),
                "total_equity": _populated_number(row.get("total_equity"), source_ref, "$m", period),
                "goodwill": _populated_number(row.get("goodwill"), source_ref, "$m", period),
                "intangibles": _populated_number(row.get("intangibles"), source_ref, "$m", period),
                "shares_outstanding": _populated_share_count(row.get("shares_outstanding"), source_ref, period),
                "pension_obligation_net": _populated_number(row.get("pension_obligation_net"), source_ref, "$m", period),
                "total_debt": _quality_checked_number(
                    value=row.get("total_debt"),
                    metric="total_debt",
                    row=row,
                    period=period,
                    history_source_ref=source_ref,
                    placeholders=placeholders,
                    review_flags=review_flags,
                ),
                "debt_core": _quality_checked_number(
                    value=row.get("debt_core"),
                    metric="debt_core",
                    row=row,
                    period=period,
                    history_source_ref=source_ref,
                    placeholders=placeholders,
                    review_flags=review_flags,
                ),
                "lease_liabilities": _populated_number(row.get("lease_liabilities"), source_ref, "$m", period),
                "lease_liabilities_current": _populated_number(row.get("lease_liabilities_current"), source_ref, "$m", period),
                "lease_liabilities_noncurrent": _populated_number(row.get("lease_liabilities_noncurrent"), source_ref, "$m", period),
                "revolver_availability": _field(
                    revolver_by_period[period],
                    source_ref=f"{workbook_path.name}!Valuation!row:95",
                    unit="$m",
                    period=period,
                )
                if period in revolver_by_period
                else _missing("Legacy revolver-availability series has no value for this period.", source_ref=f"{workbook_path.name}!Valuation!row:95"),
            }
        )
    _attach_financial_definitions(out)
    return out


def _build_annual_financial_rows(
    history_rows: Sequence[Mapping[str, Any]],
    workbook_path: Path,
    *,
    incomplete_candidates: list[dict[str, Any]],
    unsupported_zero_placeholders: Mapping[tuple[str, str], Mapping[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    by_year: dict[Any, list[Mapping[str, Any]]] = defaultdict(list)
    for row in history_rows:
        year = row.get("fiscal_year")
        if _is_present(year):
            by_year[year].append(row)
    annuals: list[dict[str, Any]] = []
    adjusted_ebitda_by_period = _read_legacy_valuation_series(workbook_path, 24)
    placeholders = unsupported_zero_placeholders or {}
    for year, rows in by_year.items():
        rows = sorted(rows, key=lambda row: _to_iso(row.get("quarter")))
        if not any(_fiscal_quarter(row) == 4 for row in rows):
            present_quarters = sorted({quarter for row in rows if (quarter := _fiscal_quarter(row)) is not None})
            missing_quarters = [quarter for quarter in range(1, 5) if quarter not in present_quarters]
            missing_labels = [f"Q{quarter}" for quarter in missing_quarters]
            source_refs = sorted({_source_ref("History_Q", row, workbook_path=workbook_path) for row in rows})
            incomplete_candidates.append(
                {
                    "period": f"{year}-FY",
                    "status": "missing_source",
                    "present_quarters": [f"Q{quarter}" for quarter in present_quarters],
                    "missing_quarters": missing_labels,
                    "source_refs": source_refs,
                    "reason": (
                        "Annual aggregation requires exactly one source-backed Q1-Q4 component; "
                        f"missing {', '.join(missing_labels)}."
                    ),
                }
            )
            continue
        source_ref = f"{workbook_path.name}!History_Q!fiscal_year:{year}"
        period = f"{year}-FY"
        components = {
            "revenue": [_annual_component(row, "revenue", unit="$", source_ref=source_ref) for row in rows],
            "cost_of_goods_sold": [_annual_component(row, "cogs", unit="$", source_ref=source_ref) for row in rows],
            "gross_profit": [_annual_component(row, "gross_profit", unit="$", source_ref=source_ref) for row in rows],
            "operating_income": [_annual_component(row, "op_income", unit="$", source_ref=source_ref) for row in rows],
            "base_ebitda": [_annual_component(row, "ebitda", unit="$", source_ref=source_ref) for row in rows],
            "adjusted_ebitda": [
                _annual_component(
                    row,
                    "adjusted_ebitda",
                    value=adjusted_ebitda_by_period.get(str(row.get("fiscal_label") or "")),
                    unit="$m",
                    source_ref=f"{workbook_path.name}!Valuation!row:24",
                )
                for row in rows
            ],
            "net_income": [_annual_component(row, "net_income", unit="$", source_ref=source_ref) for row in rows],
            "operating_cash_flow": [_annual_component(row, "cfo", unit="$", source_ref=source_ref) for row in rows],
            "capital_expenditures": [_annual_component(row, "capex", unit="$", source_ref=source_ref) for row in rows],
            "income_taxes_paid": [_annual_component(row, "tax_paid", unit="$", source_ref=source_ref) for row in rows],
            "depreciation_amortization": [_annual_component(row, "da", unit="$", source_ref=source_ref) for row in rows],
            "free_cash_flow": [
                _annual_component(
                    row,
                    "free_cash_flow",
                    value=(
                        float(row["cfo"]) - float(row["capex"])
                        if _is_present(row.get("cfo")) and _is_present(row.get("capex"))
                        else None
                    ),
                    unit="$",
                    source_ref=source_ref,
                )
                for row in rows
            ],
            "interest_paid": [
                _annual_component(
                    row,
                    "interest_paid",
                    unit="$",
                    source_ref=source_ref,
                    normalized_metric="interest_paid",
                    unsupported_zero_placeholders=placeholders,
                )
                for row in rows
            ],
            "interest_expense": [_annual_component(row, "interest_expense_net", unit="$", source_ref=source_ref) for row in rows],
            "buybacks_cash": [_annual_component(row, "buybacks_cash", unit="$", source_ref=source_ref) for row in rows],
            "dividends_cash": [_annual_component(row, "dividends_cash", unit="$", source_ref=source_ref) for row in rows],
            "acquisitions_cash": [_annual_component(row, "acquisitions_cash", unit="$", source_ref=source_ref) for row in rows],
            "debt_repayment": [_annual_component(row, "debt_repayment", unit="$", source_ref=source_ref) for row in rows],
            "debt_issuance": [_annual_component(row, "debt_issuance", unit="$", source_ref=source_ref) for row in rows],
        }
        adjusted_source_ref = f"{workbook_path.name}!Valuation!row:24"
        annuals.append(
            {
                "period": period,
                "fiscal_year": year,
                "revenue": _annual_component_field(components["revenue"], metric="revenue", period=period, source_ref=source_ref, divisor=1_000_000, core=True),
                "cost_of_goods_sold": _annual_component_field(components["cost_of_goods_sold"], metric="cost_of_goods_sold", period=period, source_ref=source_ref, divisor=1_000_000),
                "gross_profit": _annual_component_field(components["gross_profit"], metric="gross_profit", period=period, source_ref=source_ref, divisor=1_000_000),
                "operating_income": _annual_component_field(components["operating_income"], metric="operating_income", period=period, source_ref=source_ref, divisor=1_000_000, core=True),
                "base_ebitda": _annual_component_field(components["base_ebitda"], metric="base_ebitda", period=period, source_ref=source_ref, divisor=1_000_000),
                "adjusted_ebitda": _annual_component_field(components["adjusted_ebitda"], metric="adjusted_ebitda", period=period, source_ref=adjusted_source_ref),
                "net_income": _annual_component_field(components["net_income"], metric="net_income", period=period, source_ref=source_ref, divisor=1_000_000),
                "operating_cash_flow": _annual_component_field(components["operating_cash_flow"], metric="operating_cash_flow", period=period, source_ref=source_ref, divisor=1_000_000),
                "capital_expenditures": _annual_component_field(components["capital_expenditures"], metric="capital_expenditures", period=period, source_ref=source_ref, divisor=1_000_000),
                "income_taxes_paid": _annual_component_field(components["income_taxes_paid"], metric="income_taxes_paid", period=period, source_ref=source_ref, divisor=1_000_000),
                "depreciation_amortization": _annual_component_field(components["depreciation_amortization"], metric="depreciation_amortization", period=period, source_ref=source_ref, divisor=1_000_000),
                "free_cash_flow": _annual_component_field(components["free_cash_flow"], metric="free_cash_flow", period=period, source_ref=source_ref, divisor=1_000_000, core=True),
                "interest_paid": _annual_component_field(components["interest_paid"], metric="interest_paid", period=period, source_ref=source_ref, divisor=1_000_000),
                "interest_expense": _annual_component_field(components["interest_expense"], metric="interest_expense", period=period, source_ref=source_ref, divisor=1_000_000),
                "buybacks_cash": _annual_component_field(components["buybacks_cash"], metric="buybacks_cash", period=period, source_ref=source_ref, divisor=1_000_000),
                "dividends_cash": _annual_component_field(components["dividends_cash"], metric="dividends_cash", period=period, source_ref=source_ref, divisor=1_000_000),
                "acquisitions_cash": _annual_component_field(components["acquisitions_cash"], metric="acquisitions_cash", period=period, source_ref=source_ref, divisor=1_000_000),
                "debt_repayment": _annual_component_field(components["debt_repayment"], metric="debt_repayment", period=period, source_ref=source_ref, divisor=1_000_000),
                "debt_issuance": _annual_component_field(components["debt_issuance"], metric="debt_issuance", period=period, source_ref=source_ref, divisor=1_000_000),
                "cash": _annual_endpoint_field(rows, "cash", period=period, workbook_path=workbook_path),
                "marketable_securities": _annual_endpoint_field(rows, "marketable_securities", period=period, workbook_path=workbook_path),
                "total_assets": _annual_endpoint_field(rows, "assets", period=period, workbook_path=workbook_path),
                "total_liabilities": _annual_endpoint_field(rows, "liabilities", period=period, workbook_path=workbook_path),
                "current_assets": _annual_endpoint_field(rows, "assets_current", period=period, workbook_path=workbook_path),
                "current_liabilities": _annual_endpoint_field(rows, "liabilities_current", period=period, workbook_path=workbook_path),
                "inventory": _annual_endpoint_field(rows, "inventory", period=period, workbook_path=workbook_path),
                "accounts_payable": _annual_endpoint_field(rows, "accounts_payable_current", period=period, workbook_path=workbook_path),
                "accrued_liabilities": _annual_endpoint_field(rows, "accrued_liabilities_current", period=period, workbook_path=workbook_path),
                "lease_liabilities_current": _annual_endpoint_field(rows, "lease_liabilities_current", period=period, workbook_path=workbook_path),
                "lease_liabilities_noncurrent": _annual_endpoint_field(rows, "lease_liabilities_noncurrent", period=period, workbook_path=workbook_path),
                "other_assets_noncurrent": _annual_endpoint_field(rows, "other_assets_noncurrent", period=period, workbook_path=workbook_path),
                "other_liabilities_noncurrent": _annual_endpoint_field(rows, "other_liabilities_noncurrent", period=period, workbook_path=workbook_path),
                "property_plant_equipment_net": _annual_endpoint_field(rows, "property_plant_equipment_net", period=period, workbook_path=workbook_path),
                "total_equity": _annual_endpoint_field(rows, "total_equity", period=period, workbook_path=workbook_path),
                "debt_current": _annual_endpoint_field(rows, "debt_current", period=period, workbook_path=workbook_path),
                "total_debt": _annual_endpoint_field(
                    rows,
                    "total_debt",
                    period=period,
                    workbook_path=workbook_path,
                    normalized_metric="total_debt",
                    unsupported_zero_placeholders=placeholders,
                ),
                "debt_core": _annual_endpoint_field(
                    rows,
                    "debt_core",
                    period=period,
                    workbook_path=workbook_path,
                    normalized_metric="debt_core",
                    unsupported_zero_placeholders=placeholders,
                ),
                "lease_liabilities": _annual_endpoint_field(rows, "lease_liabilities", period=period, workbook_path=workbook_path),
                # A fiscal-Q4 weighted-average diluted share count is not an
                # annual EPS denominator. Preserve it explicitly for audit, but
                # do not present it as an annual diluted-share fact.
                "q4_diluted_shares": _annual_endpoint_field(rows, "shares_diluted", period=period, workbook_path=workbook_path, share_count=True),
                "diluted_shares": _missing(
                    "Annual weighted-average diluted shares are unavailable; fiscal-Q4 weighted-average shares are retained separately for audit.",
                    source_ref=f"{workbook_path.name}!History_Q!{period}",
                ),
                "shares_outstanding": _annual_endpoint_field(rows, "shares_outstanding", period=period, workbook_path=workbook_path, share_count=True),
                "eps": _missing(
                    "Source-backed annual GAAP diluted EPS or annual weighted-average diluted shares are unavailable.",
                    source_ref=f"{workbook_path.name}!History_Q!{period}",
                ),
            }
        )
    annuals.sort(key=lambda row: str(row.get("fiscal_year")))
    _attach_financial_definitions(annuals)
    return annuals


def _attach_financial_definitions(rows: Sequence[dict[str, Any]]) -> None:
    for row in rows:
        for metric, definition in FINANCIAL_FIELD_DEFINITIONS.items():
            field = row.get(metric)
            if isinstance(field, dict):
                field.setdefault("definition", definition)


def _annual_endpoint_field(
    rows: Sequence[Mapping[str, Any]],
    source_field: str,
    *,
    period: str,
    workbook_path: Path,
    share_count: bool = False,
    normalized_metric: str = "",
    unsupported_zero_placeholders: Mapping[tuple[str, str], Mapping[str, Any]] | None = None,
) -> dict[str, Any]:
    q4_rows = [row for row in rows if _fiscal_quarter(row) == 4]
    if len(q4_rows) != 1:
        return _missing(
            f"Annual endpoint {source_field} requires exactly one fiscal Q4 row; found {len(q4_rows)}.",
            source_ref=f"{workbook_path.name}!History_Q!{period}",
        )
    row = q4_rows[0]
    source_ref = _source_ref("History_Q", row, workbook_path=workbook_path)
    source_value, source_ref = _anf_history_source_evidence(row, source_field, source_ref)
    placeholder = _unsupported_zero_placeholder(
        unsupported_zero_placeholders or {},
        normalized_metric or source_field,
        row,
        row.get(source_field),
    )
    if placeholder is not None:
        field = _missing_placeholder_field(
            placeholder=placeholder,
            history_source_ref=source_ref,
            unit="m shares" if share_count else "$m",
            period=period,
        )
        field["missing_inputs"] = [str(row.get("fiscal_label") or f"{period}-Q4")]
        field["component_issues"] = [
            {
                "reason": "unsupported_zero_placeholder",
                "quarter": 4,
                "period": str(row.get("fiscal_label") or ""),
                "source_ref": str(field["source_ref"]),
                "candidate_value": 0.0,
            }
        ]
        return field
    if share_count:
        return _populated_share_count(source_value, source_ref, period)
    return _populated_number(source_value, source_ref, "$m", period)


def _annual_component_field(
    components: Sequence[Mapping[str, Any] | tuple[str, Any]],
    *,
    metric: str,
    period: str,
    source_ref: str,
    divisor: float = 1.0,
    core: bool = False,
) -> dict[str, Any]:
    expected_year_match = re.fullmatch(r"(20\d{2})-FY", period)
    expected_year = int(expected_year_match.group(1)) if expected_year_match else None
    normalized = [_normalize_annual_component(component, default_source_ref=source_ref) for component in components]
    component_issues: list[dict[str, Any]] = []
    by_quarter: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for component in normalized:
        quarter = component.get("fiscal_quarter")
        component_year = component.get("fiscal_year")
        if quarter not in {1, 2, 3, 4}:
            component_issues.append({"reason": "invalid_quarter", "component": component})
            continue
        if expected_year is not None and component_year != expected_year:
            component_issues.append(
                {
                    "reason": "mismatched_fiscal_year",
                    "expected": expected_year,
                    "actual": component_year,
                    "quarter": quarter,
                }
            )
        quality_issue = component.get("quality_issue")
        if isinstance(quality_issue, Mapping):
            component_issues.append(dict(quality_issue))
        by_quarter[int(quarter)].append(component)
    missing_inputs: list[str] = []
    selected: list[dict[str, Any]] = []
    expected_unit = "$m" if divisor == 1.0 else "$"
    for quarter in range(1, 5):
        quarter_rows = by_quarter.get(quarter, [])
        if not quarter_rows:
            missing_inputs.append(f"{expected_year or 'FY'}-Q{quarter}")
            continue
        if len(quarter_rows) != 1:
            component_issues.append(
                {
                    "reason": "duplicate_quarter",
                    "quarter": quarter,
                    "values": [row.get("value") for row in quarter_rows],
                    "source_refs": [row.get("source_ref") for row in quarter_rows],
                }
            )
            continue
        component = quarter_rows[0]
        if component.get("unit") != expected_unit:
            component_issues.append(
                {
                    "reason": "mismatched_unit",
                    "quarter": quarter,
                    "expected": expected_unit,
                    "actual": component.get("unit"),
                }
            )
        if not _is_present(component.get("value")) or component.get("status") != "populated" or not component.get("source_ref"):
            missing_inputs.append(str(component.get("label") or f"{expected_year or 'FY'}-Q{quarter}"))
        selected.append(component)
    if missing_inputs or component_issues or len(selected) != 4:
        detail_parts = []
        if missing_inputs:
            detail_parts.append("missing: " + ", ".join(missing_inputs))
        if component_issues:
            detail_parts.append("conflicts: " + ", ".join(str(item.get("reason") or "component_conflict") for item in component_issues))
        reason = f"Annual {metric} requires exactly one compatible, source-backed Q1-Q4 component; " + "; ".join(detail_parts) + "."
        field = _field(
            None,
            status="manual_review_required" if component_issues else "missing_source",
            source_ref=source_ref,
            core=core,
            reason=reason,
            unit="$m",
            period=period,
            confidence="",
        )
        field["missing_inputs"] = missing_inputs
        field["component_issues"] = component_issues
        field["component_source_refs"] = sorted({str(row.get("source_ref") or "") for row in normalized if row.get("source_ref")})
        return field
    total = sum(float(component["value"]) for component in selected)
    return _field(round(total / divisor, 3), source_ref=source_ref, core=core, unit="$m", period=period)


def _annual_component(
    row: Mapping[str, Any],
    value_field: str,
    *,
    value: Any = ...,
    unit: str,
    source_ref: str = "",
    normalized_metric: str = "",
    unsupported_zero_placeholders: Mapping[tuple[str, str], Mapping[str, Any]] | None = None,
) -> dict[str, Any]:
    label = str(row.get("fiscal_label") or "")
    actual_value = row.get(value_field) if value is ... else value
    component = {
        "label": label,
        "fiscal_year": int(row.get("fiscal_year")) if _is_present(row.get("fiscal_year")) else None,
        "fiscal_quarter": _fiscal_quarter(row),
        "value": actual_value,
        "unit": unit,
        "status": "populated" if _is_present(actual_value) else "missing_source",
        "source_ref": source_ref or str(row.get("source_ref") or f"History_Q!{label}"),
    }
    placeholder = _unsupported_zero_placeholder(
        unsupported_zero_placeholders or {},
        normalized_metric or value_field,
        row,
        actual_value,
    )
    if placeholder is not None:
        component["value"] = None
        component["status"] = "missing_source"
        component["source_ref"] = _placeholder_source_ref(str(component["source_ref"]), placeholder)
        component["quality_issue"] = {
            "reason": "unsupported_zero_placeholder",
            "quarter": component["fiscal_quarter"],
            "period": label,
            "source_ref": component["source_ref"],
            "candidate_value": 0.0,
            "source_status": str(placeholder.get("source_status") or ""),
            "qa_status": str(placeholder.get("qa_status") or ""),
        }
    return component


def _normalize_annual_component(
    component: Mapping[str, Any] | tuple[str, Any],
    *,
    default_source_ref: str,
) -> dict[str, Any]:
    if isinstance(component, Mapping):
        return dict(component)
    label, value = component
    match = re.fullmatch(r"(20\d{2})-Q([1-4])", str(label or ""))
    return {
        "label": str(label or ""),
        "fiscal_year": int(match.group(1)) if match else None,
        "fiscal_quarter": int(match.group(2)) if match else None,
        "value": value,
        "unit": "$m",
        "status": "populated" if _is_present(value) else "missing_source",
        "source_ref": default_source_ref,
    }


def _fiscal_quarter(row: Mapping[str, Any]) -> int | None:
    explicit = row.get("fiscal_quarter")
    if isinstance(explicit, int) and explicit in {1, 2, 3, 4}:
        return explicit
    match = re.search(r"-Q([1-4])$", str(row.get("fiscal_label") or ""))
    return int(match.group(1)) if match else None


def _annual_missing_component_reviews(annual_rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    reviews: list[dict[str, Any]] = []
    for row in annual_rows:
        period = str(row.get("period") or "")
        for metric, value in row.items():
            if (
                not isinstance(value, Mapping)
                or value.get("status") not in {"missing_source", "manual_review_required"}
                or not (value.get("missing_inputs") or value.get("component_issues"))
            ):
                continue
            reviews.append(
                {
                    "severity": "P2",
                    "rule_id": "legacy_adapter_annual_component_missing",
                    "issue_type": "actionable_exception",
                    "section": "annual_financials",
                    "field": f"annual_financials.rows.{period}.{metric}",
                    "normalized_path": f"annual_financials.rows.{period}.{metric}",
                    "row_key": period,
                    "affected_period": period,
                    "message": str(value.get("reason") or "Annual source components are incomplete."),
                    "source_ref": str(value.get("source_ref") or ""),
                    "missing_inputs": list(value.get("missing_inputs") or []),
                    "component_issues": list(value.get("component_issues") or []),
                    "suggested_action": "Provide every missing quarterly component before treating the annual metric as populated.",
                }
            )
    return reviews


def _annual_incomplete_candidate_reviews(
    incomplete_candidates: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    return [
        {
            "severity": "P2",
            "rule_id": "legacy_adapter_annual_fiscal_year_incomplete",
            "field": f"annual_financials.incomplete_candidates.{index}",
            "period": str(candidate.get("period") or ""),
            "message": (
                f"{candidate.get('period')}: annual source coverage is incomplete; "
                f"missing {', '.join(str(value) for value in candidate.get('missing_quarters') or [])}."
            ),
            "source_ref": " + ".join(str(value) for value in candidate.get("source_refs") or []),
            "suggested_action": "Resolve the missing fiscal-quarter evidence before deriving an annual total.",
            "visibility_disposition": "needs_review",
            "root_cause": "incomplete_annual_quarter_coverage",
            "adapter_metadata": {
                "present_quarters": list(candidate.get("present_quarters") or []),
                "missing_quarters": list(candidate.get("missing_quarters") or []),
            },
        }
        for index, candidate in enumerate(incomplete_candidates)
    ]


def _populated_number(value: Any, source_ref: str, unit: str, period: str, *, core: bool = False) -> dict[str, Any]:
    converted = _to_millions(value)
    if converted is None:
        return _missing("Legacy artifact did not contain a numeric value.", source_ref=source_ref, core=core)
    return _field(converted, source_ref=source_ref, core=core, unit=unit, period=period)


def _populated_scalar(value: Any, source_ref: str, unit: str, period: str, *, core: bool = False) -> dict[str, Any]:
    if value in (None, ""):
        return _missing("Legacy artifact did not contain a scalar value.", source_ref=source_ref, core=core)
    return _field(value, source_ref=source_ref, core=core, unit=unit, period=period)


def _populated_share_count(value: Any, source_ref: str, period: str, *, core: bool = False) -> dict[str, Any]:
    converted = _to_millions(value)
    if converted is None:
        return _missing("Legacy artifact did not contain a share count.", source_ref=source_ref, core=core)
    return _field(converted, source_ref=source_ref, core=core, unit="m shares", period=period)


def _build_debt_liquidity(
    history_rows: Sequence[Mapping[str, Any]],
    quarterly_financial_rows: Sequence[Mapping[str, Any]],
    leverage_rows: Sequence[Mapping[str, Any]],
    revolver_rows: Sequence[Mapping[str, Any]],
    workbook_path: Path,
    review_flags: list[dict[str, Any]],
) -> dict[str, Any]:
    history = next((row for row in reversed(history_rows) if _is_present(row.get("cash"))), {})
    leverage = next((row for row in reversed(leverage_rows) if _is_present(row.get("cash")) or _is_present(row.get("liquidity"))), {})
    source_ref = _source_ref("Leverage_Liquidity", leverage, workbook_path=workbook_path) if leverage else _source_ref("History_Q", history, workbook_path=workbook_path)
    cash = _to_millions(leverage.get("cash") if leverage else history.get("cash"))
    cash_period = _to_iso(leverage.get("quarter") if leverage else history.get("quarter"))
    history_period = _to_iso(history.get("quarter"))
    quality_checked_history = next(
        (
            row
            for row in reversed(quarterly_financial_rows)
            if _to_iso(row.get("period_end")) == history_period
        ),
        {},
    )
    total_debt_field = (
        quality_checked_history.get("total_debt")
        if isinstance(quality_checked_history.get("total_debt"), Mapping)
        else {}
    )
    total_debt_value = (
        total_debt_field.get("value")
        if str(total_debt_field.get("status") or "") == "populated"
        else None
    )
    total_debt = (
        float(total_debt_value)
        if isinstance(total_debt_value, (int, float)) and not isinstance(total_debt_value, bool)
        else None
    )
    debt_source_ref = str(
        total_debt_field.get("source_ref")
        or _source_ref("History_Q", history, workbook_path=workbook_path)
    )
    if total_debt is None:
        review_flags.append(
            {
                "severity": "P2",
                "rule_id": "legacy_adapter_total_debt_missing",
                "field": "debt_liquidity.total_debt",
                "message": "Latest ANF legacy history row has no source-backed total debt value; zero was not assumed.",
                "source_ref": debt_source_ref,
                "suggested_action": "Resolve total debt from source-native debt evidence before promotion.",
                "visibility_disposition": "needs_review",
            }
        )

    # Valuation!D198 is a calculated legacy display cell, not independent debt
    # evidence. Net debt remains unavailable until both debt and cash have
    # source-backed compatible lineage.
    net_debt = round(total_debt - cash, 3) if total_debt is not None and cash is not None else None
    net_debt_source_ref = f"{debt_source_ref} + {source_ref}" if net_debt is not None else debt_source_ref

    revolver = next((row for row in reversed(revolver_rows) if _is_present(row.get("revolver_availability"))), {})
    revolver_period = _to_iso(revolver.get("quarter"))
    revolver_source_ref = _source_ref("Revolver_History", revolver, workbook_path=workbook_path)
    revolver_availability = _to_millions(revolver.get("revolver_availability")) if revolver else None
    matching_cash_row = next(
        (row for row in reversed(history_rows) if _to_iso(row.get("quarter")) == revolver_period and _is_present(row.get("cash"))),
        {},
    )
    liquidity_cash = _to_millions(matching_cash_row.get("cash")) if matching_cash_row else None
    total_liquidity = (
        round(liquidity_cash + revolver_availability, 3)
        if liquidity_cash is not None and revolver_availability is not None
        else None
    )
    liquidity_source_ref = (
        f"{_source_ref('History_Q', matching_cash_row, workbook_path=workbook_path)} + {revolver_source_ref}"
        if matching_cash_row and revolver
        else revolver_source_ref or source_ref
    )
    if total_liquidity is None:
        freshness_disposition = "incomplete_components"
        freshness_reason = "A same-date cash and revolver pair is unavailable."
        summary_liquidity_display = _missing(
            "SUMMARY liquidity requires a complete same-date liquidity total.",
            source_ref=liquidity_source_ref,
            core=True,
        )
        summary_liquidity_as_of_display = _missing(
            "SUMMARY liquidity as-of text requires a complete same-date liquidity total.",
            source_ref=liquidity_source_ref,
            core=True,
        )
    elif cash_period == revolver_period:
        freshness_disposition = "current"
        freshness_reason = "Liquidity and the latest SUMMARY point-in-time evidence share one as-of date."
        summary_liquidity_display = _field(
            total_liquidity,
            source_ref=liquidity_source_ref,
            core=True,
            unit="$m",
            period=revolver_period,
        )
        summary_liquidity_as_of_display = _field(
            f"As of {revolver_period}",
            source_ref=liquidity_source_ref,
            core=True,
            period=revolver_period,
        )
    else:
        freshness_disposition = "stale_but_displayable_with_date"
        freshness_reason = "Liquidity is older than the latest SUMMARY point-in-time evidence and must be visibly dated."
        summary_liquidity_display = _field(
            total_liquidity,
            source_ref=liquidity_source_ref,
            core=True,
            unit="$m",
            period=revolver_period,
        )
        summary_liquidity_as_of_display = _field(
            f"As of {revolver_period} (stale)",
            source_ref=liquidity_source_ref,
            core=True,
            period=revolver_period,
        )
    if total_liquidity is not None and cash_period and revolver_period and cash_period != revolver_period:
        review_flags.append(
            {
                "severity": "P2",
                "rule_id": "legacy_adapter_liquidity_as_of_lag",
                "field": "debt_liquidity.total_liquidity",
                "message": f"Total liquidity is source-backed as of {revolver_period}; newer cash evidence exists as of {cash_period} without matching revolver availability.",
                "source_ref": liquidity_source_ref,
                "suggested_action": "Refresh revolver availability before treating total liquidity as current.",
                "visibility_disposition": "needs_review",
            }
        )
    return {
        "cash": _field(cash, source_ref=source_ref, core=True, unit="$m", period=cash_period) if cash is not None else _missing("Cash not found in History_Q or Leverage_Liquidity.", source_ref=source_ref, core=True),
        "total_debt": _field(total_debt, source_ref=debt_source_ref, core=True, unit="$m", period=cash_period) if total_debt is not None else _missing("Latest total debt is not source-backed; zero was not assumed.", source_ref=debt_source_ref, core=True),
        "net_debt": _field(
            net_debt,
            source_ref=net_debt_source_ref,
            core=True,
            unit="$m",
            period=cash_period,
            definition=FINANCIAL_FIELD_DEFINITIONS["net_debt"],
        )
        if net_debt is not None
        else _missing(
            "Net debt is unavailable because source-backed total debt is missing; legacy Valuation!D198 was not treated as evidence.",
            source_ref=net_debt_source_ref,
            core=True,
        ),
        "net_leverage": _missing("Net leverage cannot be source-backed while total debt is unavailable.", source_ref=debt_source_ref, core=False),
        "revolver_availability": _field(revolver_availability, source_ref=revolver_source_ref, core=True, unit="$m", period=revolver_period) if revolver_availability is not None else _missing("Revolver availability is unavailable.", source_ref=revolver_source_ref, core=True),
        "liquidity_cash": _field(liquidity_cash, source_ref=_source_ref("History_Q", matching_cash_row, workbook_path=workbook_path), core=True, unit="$m", period=revolver_period) if liquidity_cash is not None else _missing("Cash is unavailable for the total-liquidity as-of date.", source_ref=liquidity_source_ref, core=True),
        "other_available_liquidity": _not_applicable("No other available-liquidity component was normalized by the ANF legacy adapter.", source_ref=liquidity_source_ref),
        "total_liquidity": _field(total_liquidity, source_ref=liquidity_source_ref, core=True, unit="$m", period=revolver_period) if total_liquidity is not None else _missing("Total liquidity requires cash and revolver availability for the same as-of date.", source_ref=liquidity_source_ref, core=True),
        "liquidity_definition": _field("Cash plus undrawn revolver availability; excludes unverified other liquidity sources.", source_ref=liquidity_source_ref, core=True),
        "as_of_date": _field(revolver_period, source_ref=liquidity_source_ref, core=True) if revolver_period else _missing("Total-liquidity as-of date is unavailable.", source_ref=liquidity_source_ref, core=True),
        "summary_as_of_date": _field(cash_period, source_ref=source_ref, core=True) if cash_period else _missing("SUMMARY point-in-time as-of date is unavailable.", source_ref=source_ref, core=True),
        "summary_liquidity_display": summary_liquidity_display,
        "summary_liquidity_as_of_display": summary_liquidity_as_of_display,
        "liquidity_freshness": {
            "disposition": freshness_disposition,
            "summary_as_of": cash_period or "",
            "liquidity_as_of": revolver_period or "",
            "component_as_of": {
                "cash": revolver_period or "",
                "revolver": revolver_period or "",
            },
            "mixed_date_components": False,
            "reason": freshness_reason,
            "source_ref": liquidity_source_ref,
        },
        "liquidity": _field(total_liquidity, source_ref=liquidity_source_ref, unit="$m", period=revolver_period) if total_liquidity is not None else _missing("Deprecated liquidity alias remains missing because total liquidity is unavailable.", source_ref=liquidity_source_ref),
        "lease_liabilities": _populated_number(history.get("lease_liabilities"), _source_ref("History_Q", history, workbook_path=workbook_path), "$m", history_period),
        "interest_expense": _populated_number(leverage.get("interest_expense_net_ttm"), source_ref, "$m", "ttm"),
        "maturity_schedule": [],
    }


def _build_capital_returns(history_rows: Sequence[Mapping[str, Any]], workbook_path: Path) -> dict[str, Any]:
    latest_rows: list[Mapping[str, Any]] = []
    for row in sorted([item for item in history_rows if _is_present(item.get("quarter"))], key=lambda item: _to_iso(item.get("quarter"))):
        latest_rows.append(row)
        if len(latest_rows) > 4:
            latest_rows.pop(0)
    source_ref = f"{workbook_path.name}!History_Q!latest_4_quarters"
    buybacks = sum(float(row.get("buybacks_cash") or 0) for row in latest_rows)
    dividends = sum(float(row.get("dividends_cash") or 0) for row in latest_rows)
    return {
        "buybacks": _field(round(buybacks / 1_000_000, 3), source_ref=source_ref, core=True, unit="$m", period="latest_4_quarters"),
        "dividends": _field(round(dividends / 1_000_000, 3), source_ref=source_ref, unit="$m", period="latest_4_quarters")
        if dividends
        else _not_applicable("ANF has no common dividend signal in the legacy artifact.", source_ref=source_ref),
        "share_issuance": _not_applicable("No equity issuance program is represented in the legacy ANF artifact.", source_ref=source_ref),
    }


def _build_valuation_inputs(
    quarterly_rows: Sequence[Mapping[str, Any]],
    debt_liquidity: Mapping[str, Any],
    workbook_path: Path,
    review_flags: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    latest_four = list(quarterly_rows[-4:])
    source_ref = f"{workbook_path.name}!History_Q!latest_4_quarters"

    def total(field_name: str, *, expected_unit: str = "$m") -> dict[str, Any]:
        result = _ttm_component_field(
            latest_four,
            metric=field_name,
            source_ref=source_ref,
            expected_unit=expected_unit,
        )
        if str(result.get("status") or "") != "populated" and review_flags is not None:
            review_flags.append(
                {
                    "severity": "P2",
                    "rule_id": "legacy_adapter_ttm_coverage_incomplete",
                    "issue_type": "actionable_exception",
                    "section": "valuation_inputs",
                    "field": f"valuation_inputs.{field_name}_ttm",
                    "normalized_path": f"valuation_inputs.{field_name}_ttm",
                    "row_key": field_name,
                    "message": str(result.get("reason") or f"TTM {field_name} coverage is incomplete."),
                    "source_ref": str(result.get("source_ref") or source_ref),
                    "suggested_action": "Provide exactly four consecutive, compatible, source-backed quarterly components.",
                    "adapter_metadata": {
                        "available_quarters": list(result.get("available_quarters") or []),
                        "expected_quarters": list(result.get("expected_quarters") or []),
                        "missing_quarters": list(result.get("missing_quarters") or []),
                        "duplicate_quarters": list(result.get("duplicate_quarters") or []),
                        "component_issues": list(result.get("component_issues") or []),
                    },
                }
            )
        return result

    latest = latest_four[-1] if latest_four else {}
    latest_source = str(latest.get("revenue", {}).get("source_ref") or source_ref) if isinstance(latest.get("revenue"), Mapping) else source_ref
    as_of = str(latest.get("period_end") or "")
    diluted = latest.get("diluted_shares") if isinstance(latest.get("diluted_shares"), Mapping) else _missing("Latest diluted shares unavailable.", source_ref=latest_source, core=True)
    outstanding = latest.get("shares_outstanding") if isinstance(latest.get("shares_outstanding"), Mapping) else _missing("Latest point-in-time shares outstanding unavailable.", source_ref=latest_source, core=True)
    equity = latest.get("total_equity") if isinstance(latest.get("total_equity"), Mapping) else {}
    goodwill = latest.get("goodwill") if isinstance(latest.get("goodwill"), Mapping) else {}
    intangibles = latest.get("intangibles") if isinstance(latest.get("intangibles"), Mapping) else {}

    def per_share(numerator: Mapping[str, Any], *, name: str) -> dict[str, Any]:
        numerator_value = numerator.get("value") if str(numerator.get("status") or "") == "populated" else None
        outstanding_value = outstanding.get("value") if isinstance(outstanding, Mapping) and str(outstanding.get("status") or "") == "populated" else None
        refs = [str(item.get("source_ref") or "") for item in (numerator, outstanding) if isinstance(item, Mapping)]
        if not isinstance(numerator_value, (int, float)) or isinstance(numerator_value, bool) or not isinstance(outstanding_value, (int, float)) or isinstance(outstanding_value, bool) or outstanding_value == 0:
            return _missing(f"{name} requires source-backed equity and point-in-time shares outstanding.", source_ref=" + ".join(filter(None, refs)), core=True)
        definition_key = "tangible_book_value_per_share" if name.startswith("Tangible") else "book_value_per_share"
        return _field(
            round(float(numerator_value) / float(outstanding_value), 4),
            source_ref=" + ".join(filter(None, refs)),
            core=True,
            unit="$/share",
            period=as_of,
            definition=FINANCIAL_FIELD_DEFINITIONS[definition_key],
        )

    book_value_per_share = per_share(equity, name="Book value per share")
    tangible_value: dict[str, Any]
    if all(str(field.get("status") or "") == "populated" for field in (equity, goodwill, intangibles)):
        tangible_equity = _field(
            float(equity["value"]) - float(goodwill["value"]) - float(intangibles["value"]),
            source_ref=" + ".join(str(field.get("source_ref") or "") for field in (equity, goodwill, intangibles)),
            core=True,
            unit="$m",
            period=as_of,
        )
        tangible_value = per_share(tangible_equity, name="Tangible book value per share")
    else:
        tangible_value = _missing("Tangible book value requires source-backed equity, goodwill, intangibles, and point-in-time shares outstanding; missing components were not assumed to be zero.", source_ref=latest_source, core=True)
    shares = outstanding
    return {
        "price": _missing("Market price is intentionally not sourced by the ANF legacy adapter fixture.", source_ref=f"{workbook_path.name}!Valuation!D194"),
        "as_of_date": _field(as_of, source_ref=latest_source, core=True) if as_of else _missing("Latest period end is unavailable.", source_ref=latest_source, core=True),
        "shares_outstanding": shares,
        "diluted_shares": diluted,
        "net_debt": dict(debt_liquidity.get("net_debt") or _missing("Net debt unavailable.", source_ref=source_ref, core=True)),
        "base_ebitda_ttm": total("base_ebitda"),
        "adjusted_ebitda_ttm": total("adjusted_ebitda"),
        "revenue_ttm": total("revenue"),
        "net_income_ttm": total("net_income"),
        "operating_cash_flow_ttm": total("operating_cash_flow"),
        "free_cash_flow_ttm": total("free_cash_flow"),
        "capex_ttm": total("capital_expenditures"),
        "interest_paid_ttm": total("interest_paid"),
        "eps_ttm": total("eps", expected_unit="$/share"),
        "adjusted_eps_ttm": total("adjusted_eps", expected_unit="$/share"),
        "book_value_per_share": book_value_per_share,
        "tangible_book_value_per_share": tangible_value,
        "adjusted_fcf_ttm": _missing("No independently source-backed adjusted FCF TTM definition is available in the ANF legacy fixture.", source_ref=source_ref),
        "target_ev_adjusted_ebitda": _missing("Target EV/Adjusted EBITDA is a user assumption and is intentionally blank.", source_ref="user_input"),
        "target_ev_ebitda": _missing("Target EV/EBITDA is a user assumption and is intentionally blank.", source_ref="user_input"),
        "target_ev_yield": _missing("Target EV yield is a user assumption and is intentionally blank.", source_ref="user_input"),
        "maintenance_capex_ratio": _missing("Maintenance capex ratio is a user assumption and is intentionally blank.", source_ref="user_input"),
        "recurring_cash_costs": _missing("Recurring cash costs are a user assumption and are intentionally blank.", source_ref="user_input"),
        "working_capital_normalization": _missing("Working-capital normalization is a user assumption and is intentionally blank.", source_ref="user_input"),
        "per_share_denominator": _missing("Per-share denominator is a user selection and is intentionally blank.", source_ref="user_input"),
    }


def _quarter_ordinal(period: str) -> int | None:
    match = re.fullmatch(r"(20\d{2})-Q([1-4])", str(period or "").strip())
    if not match:
        return None
    return int(match.group(1)) * 4 + int(match.group(2)) - 1


def _quarter_from_ordinal(ordinal: int) -> str:
    year, quarter_index = divmod(ordinal, 4)
    return f"{year}-Q{quarter_index + 1}"


def _build_calculation_history(quarterly_rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    """Project full source-backed quarterly inputs for formula-owned history.

    The visible workbook keeps a bounded display axis. This long-form hidden
    projection preserves older quarters needed to calculate TTM and YoY values
    for every visible period without copying legacy formula outputs.
    """

    items: list[dict[str, Any]] = []
    for row in quarterly_rows:
        period = str(row.get("period") or "")
        ordinal = _quarter_ordinal(period)
        if ordinal is None:
            continue
        for metric in CALCULATION_HISTORY_METRICS:
            field = row.get(metric)
            if not isinstance(field, Mapping) or str(field.get("status") or "") != "populated":
                continue
            value = field.get("value")
            unit = str(field.get("unit") or "")
            source_ref = str(field.get("source_ref") or "")
            if not isinstance(value, (int, float)) or isinstance(value, bool) or not unit or not source_ref:
                continue
            items.append(
                {
                    "period": period,
                    "period_ordinal": ordinal,
                    "metric": metric,
                    "value": value,
                    "unit": unit,
                    "source_ref": source_ref,
                    "status": "populated",
                    "definition": str(field.get("definition") or FINANCIAL_FIELD_DEFINITIONS.get(metric) or ""),
                }
            )
    items.sort(key=lambda item: (int(item["period_ordinal"]), str(item["metric"])))
    return {"quarterly_items": items}


def _ttm_component_field(
    quarterly_rows: Sequence[Mapping[str, Any]],
    *,
    metric: str,
    source_ref: str,
    expected_end_period: str = "",
    expected_unit: str = "$m",
) -> dict[str, Any]:
    components: list[dict[str, Any]] = []
    component_issues: list[dict[str, Any]] = []
    by_period: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for source_index, row in enumerate(quarterly_rows):
        period = str(_normalized_scalar(row.get("period")) or "").strip()
        field = row.get(metric) if isinstance(row.get(metric), Mapping) else {}
        component = {
            "source_index": source_index,
            "period": period,
            "ordinal": _quarter_ordinal(period),
            "value": field.get("value"),
            "status": str(field.get("status") or ""),
            "unit": str(field.get("unit") or ""),
            "source_ref": str(field.get("source_ref") or ""),
            "dimension": _normalized_scalar(field.get("dimension")) or _normalized_scalar(row.get("dimension")),
            "member": _normalized_scalar(field.get("member")) or _normalized_scalar(row.get("member")),
        }
        components.append(component)
        if component["ordinal"] is None:
            component_issues.append({"reason": "invalid_quarter", "period": period, "source_index": source_index})
        else:
            by_period[period].append(component)

    duplicate_quarters = sorted(period for period, rows in by_period.items() if len(rows) != 1)
    for period in duplicate_quarters:
        rows = by_period[period]
        component_issues.append(
            {
                "reason": "duplicate_quarter",
                "period": period,
                "values": [row.get("value") for row in rows],
                "source_refs": [row.get("source_ref") for row in rows],
            }
        )

    ordinals = sorted({int(component["ordinal"]) for component in components if component.get("ordinal") is not None})
    expected_end_ordinal = _quarter_ordinal(expected_end_period) if expected_end_period else (ordinals[-1] if ordinals else None)
    if expected_end_period and expected_end_ordinal is None:
        component_issues.append({"reason": "invalid_expected_end_period", "period": expected_end_period})
    expected_ordinals = (
        list(range(expected_end_ordinal - 3, expected_end_ordinal + 1))
        if expected_end_ordinal is not None
        else []
    )
    expected_quarters = [_quarter_from_ordinal(value) for value in expected_ordinals]
    available_quarters = [str(component.get("period") or "") for component in components]
    missing_quarters = [period for period in expected_quarters if period not in by_period]
    if len(components) != 4:
        component_issues.append(
            {
                "reason": "quarter_count_not_four",
                "expected": 4,
                "actual": len(components),
            }
        )
    if len(ordinals) == 4 and ordinals != expected_ordinals:
        component_issues.append(
            {
                "reason": "quarters_not_consecutive",
                "actual": [_quarter_from_ordinal(value) for value in ordinals],
                "expected": expected_quarters,
            }
        )
    elif len(components) == 4 and len(ordinals) != 4:
        component_issues.append(
            {
                "reason": "quarters_not_distinct",
                "actual": available_quarters,
            }
        )

    component_units = [str(component.get("unit") or "").strip() for component in components]
    if len(component_units) == 4 and (
        any(not unit for unit in component_units) or set(component_units) != {expected_unit}
    ):
        component_issues.append(
            {
                "reason": "incompatible_unit",
                "expected": expected_unit,
                "actual": component_units,
            }
        )
    dimensions = {
        json.dumps(
            {"dimension": component.get("dimension"), "member": component.get("member")},
            sort_keys=True,
            ensure_ascii=False,
        )
        for component in components
    }
    if len(dimensions) > 1:
        component_issues.append({"reason": "incompatible_dimensions", "actual": sorted(dimensions)})

    missing_value_periods: list[str] = []
    for component in components:
        value = component.get("value")
        if (
            component.get("status") != "populated"
            or not isinstance(value, (int, float))
            or isinstance(value, bool)
            or not str(component.get("source_ref") or "").strip()
        ):
            missing_value_periods.append(str(component.get("period") or f"source_index:{component['source_index']}"))
    if missing_value_periods:
        component_issues.append(
            {
                "reason": "component_not_source_backed_numeric",
                "periods": missing_value_periods,
            }
        )

    source_refs = sorted({str(component.get("source_ref") or "") for component in components if component.get("source_ref")})
    valid = (
        len(components) == 4
        and len(ordinals) == 4
        and ordinals == expected_ordinals
        and not duplicate_quarters
        and not missing_quarters
        and not component_issues
    )
    if valid:
        return _field(
            round(sum(float(component["value"]) for component in components), 3),
            source_ref=" + ".join(source_refs) or source_ref,
            core=True,
            unit=expected_unit,
            period="TTM",
        )

    conflict_reasons = {str(issue.get("reason") or "") for issue in component_issues}
    conflict = bool(
        conflict_reasons
        & {
            "invalid_quarter",
            "invalid_expected_end_period",
            "duplicate_quarter",
            "quarters_not_consecutive",
            "quarters_not_distinct",
            "incompatible_unit",
            "incompatible_dimensions",
        }
    )
    reason = (
        f"TTM {metric} requires exactly four distinct consecutive, compatible, source-backed quarterly values; "
        + ", ".join(sorted(conflict_reasons or {"insufficient_quarter_coverage"}))
        + "."
    )
    field = _field(
        None,
        status="manual_review_required" if conflict else "missing_source",
        source_ref=" + ".join(source_refs) or source_ref,
        core=True,
        reason=reason,
        unit=expected_unit,
        period="TTM",
        confidence="",
    )
    field.update(
        {
            "available_quarters": available_quarters,
            "expected_quarters": expected_quarters,
            "missing_quarters": missing_quarters,
            "duplicate_quarters": duplicate_quarters,
            "component_issues": component_issues,
            "component_source_refs": source_refs,
        }
    )
    return field


def _typed_guidance_comparison_contract(
    row: Mapping[str, Any],
    *,
    metric: str,
    horizon: str,
    unit: str,
    source_ref: str,
) -> dict[str, Any] | None:
    """Return only an explicit single-metric range contract from structured fields."""

    low = row.get("low")
    high = row.get("high")
    if (
        not isinstance(low, (int, float))
        or isinstance(low, bool)
        or not isinstance(high, (int, float))
        or isinstance(high, bool)
        or float(low) > float(high)
        or not metric
        or not horizon
        or unit not in _NORMALIZED_UNITS
    ):
        return None
    scope = normalize_guidance_scope(
        {
            "metric": {"value": metric},
            "horizon": {"value": horizon},
            "value": {"unit": unit},
        }
    )
    if not scope.metric or not scope.horizon:
        return None
    return {
        "comparison_type": "range",
        "metric": scope.metric,
        "low": float(low),
        "high": float(high),
        "unit": unit,
        "horizon": scope.horizon,
        "source_ref": source_ref,
    }


def _anf_legacy_guidance_horizon(
    *,
    row_number: int,
    document: str,
    page: Any,
    metric: str,
    numbers: str,
    legacy_horizon: str,
) -> tuple[str, str]:
    """Correct five mislabelled FY2025 rows in the read-only ANF fixture."""

    expected = _ANF_FY2025_PRE_RELEASE_FULL_YEAR_ROWS.get(row_number)
    if expected is None:
        return legacy_horizon, ""

    try:
        source_page = int(page)
    except (TypeError, ValueError):
        source_page = 0
    observed = (Path(document).name, source_page, metric, numbers)
    required = (_ANF_FY2025_PRE_RELEASE_DOCUMENT, 1, *expected)
    if observed != required:
        raise ValueError(
            f"ANF Guidance_Normalized row {row_number} no longer matches the reviewed "
            f"FY2025 pre-release fixture contract: expected {required!r}, got {observed!r}."
        )
    return "FY2025", _ANF_FY2025_PRE_RELEASE_CONTEXT


def _build_guidance_items(
    guidance_rows: Sequence[Mapping[str, Any]],
    promise_rows: Sequence[Mapping[str, Any]],
    workbook_path: Path,
    demotions: list[dict[str, Any]],
    routing_reviews: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    clean_rows = [row for row in guidance_rows if _is_present(row.get("metric_hint")) and _is_present(row.get("numbers"))]
    clean_rows.sort(key=lambda row: (_to_iso(row.get("quarter")), str(row.get("metric_hint")), str(row.get("period_label"))))
    progress_by_key: dict[tuple[str, str], Mapping[str, Any]] = {}
    for row in promise_rows:
        key = (str(row.get("metric_display") or row.get("metric_ref") or ""), str(row.get("target_period_label") or ""))
        if key[0] and key[1]:
            progress_by_key[key] = row
    items: list[dict[str, Any]] = []
    for row in clean_rows:
        metric = str(row.get("metric") or row.get("metric_hint") or "").strip()
        legacy_horizon = str(row.get("horizon_label") or row.get("period_label") or "").strip()
        horizon, source_table_context = _anf_legacy_guidance_horizon(
            row_number=int(row.get("_row_number") or 0),
            document=str(row.get("doc") or ""),
            page=row.get("page"),
            metric=metric,
            numbers=str(row.get("numbers") or "").strip(),
            legacy_horizon=legacy_horizon,
        )
        progress = progress_by_key.get((f"{metric} guidance", horizon)) or progress_by_key.get((metric, horizon)) or {}
        source_document = str(row.get("doc") or "").strip()
        legacy_row_ref = _source_ref("Guidance_Normalized", row, workbook_path=workbook_path)
        source_ref = f"{source_document}#{legacy_row_ref}" if source_document else legacy_row_ref
        publication_date = _publication_date_from_source(source_ref, row.get("source_date") or row.get("quarter"))
        value_text = str(row.get("numbers") or row.get("value") or "").strip()
        legacy_stated_in = str(row.get("stated_in_label") or "").strip()
        source_date = _to_iso(row.get("source_date") or row.get("quarter"))
        stated_in_period = _legacy_guidance_reporting_period(
            legacy_stated_in,
            source_date=source_date,
            horizon=horizon,
        )
        source_line_raw = _clean_text(row.get("line"), limit=260)
        field_prefix = f"normalized_guidance.items.{len(items)}"
        source_line = _visible_text_or_blank(
            source_line_raw,
            field=f"{field_prefix}.source_excerpt",
            section="normalized_guidance",
            source_ref=source_ref,
            demotions=demotions,
        )
        notes_source_raw = _clean_text(progress.get("rationale") or source_line_raw, limit=220)
        notes_source = _visible_text_or_blank(
            notes_source_raw,
            field=f"{field_prefix}.notes_source",
            section="normalized_guidance",
            source_ref=_source_ref("Promise_Progress", progress, workbook_path=workbook_path) if progress else source_ref,
            demotions=demotions,
        )
        unit = str(row.get("unit") or "")
        comparison_contract = _typed_guidance_comparison_contract(
            row,
            metric=metric,
            horizon=horizon,
            unit=unit,
            source_ref=source_ref,
        )
        item = {
                "metric": _field(metric, source_ref=source_ref, core=True),
                "value": _field(value_text, source_ref=source_ref, core=True, unit=unit),
                "horizon": _field(horizon, source_ref=source_ref, core=True),
                "comparison_contract": comparison_contract,
                "comparison_contract_disposition": (
                    "typed_single_metric_range"
                    if comparison_contract is not None
                    else "manual_review_required_no_compatible_typed_range"
                ),
                "source_excerpt": source_line,
                "source_date": source_date,
                "publication_date": publication_date,
                "stated_in_period": stated_in_period,
                "legacy_stated_in_label": legacy_stated_in,
                "classification": str(row.get("source_context") or "normalized_outlook"),
                "evidence_key": _evidence_key(source_ref, metric, horizon, value_text, source_line_raw),
                "evidence_refs": [source_ref],
                "initial_guide": _missing("No distinct earlier guide was normalized for this evidence row.", source_ref=source_ref),
                "q1_update": _field("", status="missing_source", source_ref=source_ref, reason="Progression update columns are not normalized from legacy artifacts yet."),
                "q2_update": _field("", status="missing_source", source_ref=source_ref, reason="Progression update columns are not normalized from legacy artifacts yet."),
                "q3_update": _field("", status="missing_source", source_ref=source_ref, reason="Progression update columns are not normalized from legacy artifacts yet."),
                "q4_update": _field("", status="missing_source", source_ref=source_ref, reason="Progression update columns are not normalized from legacy artifacts yet."),
                "actual": _field(progress.get("actual"), source_ref=_source_ref("Promise_Progress", progress, workbook_path=workbook_path)) if progress.get("actual") not in (None, "") else _field("", status="missing_source", source_ref=_source_ref("Promise_Progress", progress, workbook_path=workbook_path) if progress else source_ref, reason="No actual/result value is normalized for this guidance row yet."),
                "progress_status": str(progress.get("status") or "open"),
                "notes_source": notes_source,
                "update_stage": "initial" if publication_date[5:7] in {"02", "03"} else "update",
                "display_role": "history",
                "display_priority": 999,
        }
        if source_table_context:
            item["source_table_context"] = source_table_context
        items.append(item)
    return _route_guidance_items(items, routing_reviews=routing_reviews)


def _guidance_source_rank(source_ref: str) -> tuple[int, str]:
    lowered = source_ref.casefold()
    if any(marker in lowered for marker in (".htm#", ".html#")) or lowered.endswith((".htm", ".html")):
        return (0, lowered)
    if ".pdf#" in lowered or lowered.endswith(".pdf"):
        return (1, lowered)
    return (2, lowered)


def _guidance_value_signature(item: Mapping[str, Any]) -> tuple[Any, ...]:
    scope = normalize_guidance_scope(item)
    value = str((item.get("value") or {}).get("value") or "") if isinstance(item.get("value"), Mapping) else str(item.get("value") or "")
    return (
        scope.scope_key,
        str(item.get("publication_date") or ""),
        str(item.get("source_date") or ""),
        str(item.get("stated_in_period") or ""),
        re.sub(r"\s+", " ", value.strip().casefold()),
    )


def _route_guidance_items(
    raw_items: Sequence[dict[str, Any]],
    *,
    routing_reviews: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Deduplicate evidence and route guidance by business scope and publication."""

    exact_groups: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for item in raw_items:
        exact_groups[_guidance_value_signature(item)].append(item)

    items: list[dict[str, Any]] = []
    for signature, rows in sorted(exact_groups.items(), key=lambda pair: str(pair[0])):
        ordered = sorted(rows, key=lambda row: _guidance_source_rank(str((row.get("value") or {}).get("source_ref") or "")))
        retained = ordered[0]
        evidence_refs = list(
            dict.fromkeys(
                str((row.get("value") or {}).get("source_ref") or "")
                for row in ordered
                if str((row.get("value") or {}).get("source_ref") or "")
            )
        )
        retained["evidence_refs"] = evidence_refs
        retained["duplicate_evidence_count"] = len(ordered)
        retained["evidence_key"] = _evidence_key("guidance_semantic", *signature)
        if len(ordered) > 1:
            retained["duplicate_evidence_disposition"] = "collapsed_with_all_source_refs_retained"
        items.append(retained)

    scope_groups: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for item in items:
        scope_groups[guidance_scope_key(item)].append(item)

    for scoped_rows in scope_groups.values():
        scoped_rows.sort(key=lambda row: (str(row.get("publication_date") or ""), str(row.get("evidence_key") or "")))
        first = scoped_rows[0]
        first_value = first.get("value") if isinstance(first.get("value"), Mapping) else _missing("Initial guidance value is unavailable.")
        previous: dict[str, Any] | None = None
        for item in scoped_rows:
            source_ref = str((item.get("value") or {}).get("source_ref") or "")
            item["display_role"] = "history"
            item["visibility_disposition"] = "historical"
            item["initial_guide"] = dict(first_value)
            item["update_stage"] = "initial" if previous is None else (
                "reaffirmed"
                if str((previous.get("value") or {}).get("value") or "") == str((item.get("value") or {}).get("value") or "")
                else "update"
            )
            if previous is not None:
                previous["superseded_by_evidence_key"] = item["evidence_key"]
                item["supersedes_evidence_keys"] = [previous["evidence_key"]]
                previous["display_role"] = "superseded"
                previous["visibility_disposition"] = "superseded"
                previous["disposition_reason"] = "superseded_by_later_same_scope_update"
            item["notes_source"] = (
                f"Published {item.get('publication_date')}; stated in {item.get('stated_in_period')}."
            )
            item["review_state"] = "accepted"
            item["source_ref"] = source_ref
            previous = item

    conflicts: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for item in items:
        conflicts[(*guidance_scope_key(item), str(item.get("publication_date") or ""))].append(item)
    for scoped_rows in conflicts.values():
        values = {
            re.sub(r"\s+", " ", str((item.get("value") or {}).get("value") or "").strip().casefold())
            for item in scoped_rows
        }
        if len(values) <= 1:
            continue
        for item in scoped_rows:
            item["display_role"] = "audit_only"
            item["visibility_disposition"] = "audit_only"
            item["disposition_reason"] = "conflicting_values_for_same_scope_and_publication"
            item["review_state"] = "manual_review_required"
            routing_reviews.append(
                {
                    "severity": "P2",
                    "rule_id": "guidance_same_publication_conflict",
                    "field": f"normalized_guidance.items.{item.get('evidence_key')}",
                    "message": "Conflicting guidance values share one normalized scope and publication date; all remain audit-only.",
                    "source_ref": str((item.get("value") or {}).get("source_ref") or ""),
                    "suggested_action": "Resolve the source context before promoting one value to visible guidance.",
                    "section": "normalized_guidance",
                    "classification": "manual_review_required",
                }
            )

    if not items:
        return []
    latest_publication = max(str(item.get("publication_date") or "") for item in items)
    current_year = latest_publication[:4]
    current_items = [
        item
        for item in items
        if item.get("display_role") != "audit_only"
        and str(item.get("publication_date") or "") == latest_publication
        and normalize_guidance_scope(item).fiscal_year == int(current_year)
    ]
    current_items.sort(
        key=lambda item: (
            normalize_guidance_scope(item).metric,
            normalize_guidance_scope(item).horizon,
            str(item.get("evidence_key") or ""),
        )
    )
    for priority, item in enumerate(current_items, start=100):
        item["display_role"] = "current_secondary"
        item["display_priority"] = priority
        item["visibility_disposition"] = "visible"
        item["disposition_reason"] = "latest_current_scope_secondary"

    pair_priority = {
        ("revenue", f"FY{current_year}"): 1,
        ("revenue", f"{current_year}-Q1"): 2,
        ("operating_margin", f"FY{current_year}"): 3,
        ("operating_margin", f"{current_year}-Q1"): 4,
        ("adjusted_eps", f"FY{current_year}"): 5,
        ("adjusted_eps", f"{current_year}-Q1"): 6,
        ("real_estate_activity", f"FY{current_year}"): 7,
    }
    for item in current_items:
        scope = normalize_guidance_scope(item)
        priority = pair_priority.get((scope.metric, scope.horizon))
        if priority is None:
            continue
        item["display_role"] = "current_primary"
        item["display_priority"] = priority
        item["visibility_disposition"] = "visible"
        item["disposition_reason"] = "latest_current_scope_primary"

    items.sort(
        key=lambda item: (
            str(item.get("publication_date") or ""),
            normalize_guidance_scope(item).metric,
            normalize_guidance_scope(item).horizon,
            str(item.get("evidence_key") or ""),
        )
    )
    return items


def _legacy_bs_segment_items(workbook_path: Path) -> list[dict[str, Any]]:
    """Inventory reliable visible legacy segment facts by exact business key."""

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb["BS_Segments"]
        items: list[dict[str, Any]] = []
        blocks = (
            ("quarterly", 7, (61, 62, 63, 65, 66, 67), range(2, 14)),
            ("annual", 70, (72, 73, 74), range(2, 10)),
        )
        display_order: dict[tuple[str, str], int] = {}
        for period_type, header_row, member_rows, columns in blocks:
            for row_number in member_rows:
                member = str(ws.cell(row_number, 1).value or "").strip()
                if not member:
                    continue
                dimension = _segment_dimension(member)
                key = (dimension, member)
                display_order.setdefault(
                    key,
                    1 + sum(1 for existing_dimension, _ in display_order if existing_dimension == dimension),
                )
                for column in columns:
                    raw_period = ws.cell(header_row, column).value
                    value = ws.cell(row_number, column).value
                    if raw_period in (None, "") or not isinstance(value, (int, float)) or isinstance(value, bool):
                        continue
                    period = (
                        f"{int(raw_period)}-FY"
                        if period_type == "annual" and isinstance(raw_period, (int, float))
                        else _normalize_period(raw_period, period_type=period_type)
                    )
                    source_ref = f"{workbook_path.name}!BS_Segments!{ws.cell(row_number, column).coordinate}"
                    items.append(
                        {
                            "dimension": dimension,
                            "member": member,
                            "display_order": display_order[key],
                            "segment": _field(member, source_ref=source_ref, core=True),
                            "metric": "revenue",
                            "period": period,
                            "period_type": period_type,
                            "unit": "$m",
                            "source_unit": "$m",
                            "source_scale": "millions",
                            "source_table_scope": period_type,
                            "source_table_id": f"{workbook_path.name}:BS_Segments:{period_type}",
                            "source_row_ref": f"BS_Segments!{ws.cell(row_number, column).coordinate}",
                            "source_ref": source_ref,
                            "aggregation_role": segment_aggregation_role(dimension, member),
                            "source": "legacy_visible_segment_oracle",
                            "note": _missing(
                                "The legacy visible segment matrix contains a numeric fact without a separate narrative note.",
                                source_ref=source_ref,
                            ),
                            "revenue": _field(
                                float(value),
                                source_ref=source_ref,
                                core=True,
                                unit="$m",
                                period=period,
                                definition="Revenue for the stated segment dimension, member, and fiscal period.",
                            ),
                        }
                    )
                    if period_type == "annual":
                        items[-1]["annual_revenue"] = dict(items[-1]["revenue"])
        return items
    finally:
        wb.close()


def _table_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _find_anf_segment_table(path: Path, *, scope_label: str, scale_label: str) -> tuple[int, Any]:
    for table_index, table in enumerate(pd.read_html(path)):
        text = " ".join(_table_text(value) for value in table.to_numpy().ravel())
        if scope_label in text and scale_label in text and "Net sales by" in text:
            return table_index, table
    raise SegmentNormalizationError(
        f"Could not locate {scope_label!r} {scale_label!r} segment table in {path}."
    )


def _table_scope_start(table: Any, scope_label: str) -> int:
    columns = [
        column
        for row in range(table.shape[0])
        for column in range(table.shape[1])
        if _table_text(table.iat[row, column]) == scope_label
    ]
    if not columns:
        raise SegmentNormalizationError(f"Table has no {scope_label!r} scope columns.")
    return min(columns)


def _table_section_rows(table: Any, section_label: str, stop_label: str | None = None) -> list[int]:
    section_row = next(
        (
            row
            for row in range(table.shape[0])
            if _table_text(table.iat[row, 0]).lower().startswith(section_label.lower())
        ),
        None,
    )
    if section_row is None:
        raise SegmentNormalizationError(f"Table has no {section_label!r} section.")
    stop_row = table.shape[0]
    if stop_label:
        stop_row = next(
            (
                row
                for row in range(section_row + 1, table.shape[0])
                if _table_text(table.iat[row, 0]).lower().startswith(stop_label.lower())
            ),
            table.shape[0],
        )
    return list(range(section_row + 1, stop_row))


def _table_member_row(table: Any, rows: Sequence[int], member_label: str) -> int:
    matches = [
        row
        for row in rows
        if _table_text(table.iat[row, 0]).lower().startswith(member_label.lower())
    ]
    if len(matches) != 1:
        raise SegmentNormalizationError(
            f"Expected one {member_label!r} row in source table, found {len(matches)}."
        )
    return matches[0]


def _table_numeric_value(table: Any, row: int, first_column: int) -> float:
    values: list[float] = []
    for value in table.iloc[row, first_column:].tolist():
        if isinstance(value, bool) or pd.isna(value):
            continue
        try:
            values.append(float(value))
        except (TypeError, ValueError):
            continue
    if not values:
        raise SegmentNormalizationError(f"Source table row {row} has no numeric value from column {first_column}.")
    return values[0]


def _anf_authoritative_segment_source_facts(data_root: Path) -> list[SegmentSourceFact]:
    source_root = data_root / "tickers" / "ANF" / "earnings_release"
    q4_path = source_root / "8-K_2024-03-07_earnings_release.htm"
    fy_path = source_root / "8-K_2019-03-07_earnings_release.htm"
    if not q4_path.exists() or not fy_path.exists():
        missing = [str(path) for path in (q4_path, fy_path) if not path.exists()]
        raise SegmentNormalizationError(f"Missing authoritative ANF segment source table(s): {missing!r}.")

    facts: list[SegmentSourceFact] = []
    q4_table_index, q4_table = _find_anf_segment_table(
        q4_path,
        scope_label="Fourth Quarter",
        scale_label="(in thousands)",
    )
    q4_start = _table_scope_start(q4_table, "Fourth Quarter")
    q4_sections = (
        (
            "Net sales by segment",
            "Net sales by brand",
            (
                ("Americas", "Americas", "geography"),
                ("EMEA", "EMEA", "geography"),
                ("APAC", "APAC", "geography"),
                ("Total company", "Total Company", "total_company"),
            ),
        ),
        (
            "Net sales by brand",
            None,
            (
                ("Hollister", "Hollister", "brand"),
                ("Abercrombie", "Abercrombie", "brand"),
            ),
        ),
    )
    for section_label, stop_label, members in q4_sections:
        section_rows = _table_section_rows(q4_table, section_label, stop_label)
        for source_label, member, dimension in members:
            row = _table_member_row(q4_table, section_rows, source_label)
            row_ref = f"table[{q4_table_index}]!row[{row}]"
            facts.append(
                SegmentSourceFact(
                    metric="revenue",
                    value=_table_numeric_value(q4_table, row, q4_start),
                    source_unit="USD",
                    source_scale="thousands",
                    period_type="quarterly",
                    period="2023-Q4",
                    dimension=dimension,
                    member=member,
                    source_table_scope="quarterly",
                    source_table_id=f"{q4_path.name}:table[{q4_table_index}]:fourth_quarter",
                    source_row_ref=row_ref,
                    source_ref=f"{q4_path}!{row_ref}",
                )
            )

    fy_table_index, fy_table = _find_anf_segment_table(
        fy_path,
        scope_label="Full Year",
        scale_label="(in millions)",
    )
    fy_start = _table_scope_start(fy_table, "Full Year")
    fy_rows = _table_section_rows(fy_table, "Net sales by brand", "Net sales by region")
    for source_label, member, dimension in (
        ("Total company", "Total Company", "total_company"),
        ("Hollister", "Hollister", "brand"),
        ("Abercrombie", "Abercrombie", "brand"),
    ):
        row = _table_member_row(fy_table, fy_rows, source_label)
        row_ref = f"table[{fy_table_index}]!row[{row}]"
        facts.append(
            SegmentSourceFact(
                metric="revenue",
                value=_table_numeric_value(fy_table, row, fy_start),
                source_unit="USD",
                source_scale="millions",
                period_type="annual",
                period="2018-FY",
                dimension=dimension,
                member=member,
                source_table_scope="annual",
                source_table_id=f"{fy_path.name}:table[{fy_table_index}]:full_year",
                source_row_ref=row_ref,
                source_ref=f"{fy_path}!{row_ref}",
            )
        )
    return list(canonicalize_segment_source_facts(facts))


def _normalized_segment_revenue_item(
    fact: SegmentSourceFact,
    *,
    display_order: int,
) -> dict[str, Any]:
    value = fact.normalized_value
    item: dict[str, Any] = {
        "dimension": fact.dimension,
        "member": fact.member,
        "display_order": display_order,
        "segment": _field(fact.member, source_ref=fact.source_ref, core=True),
        "metric": fact.metric,
        "period": fact.period,
        **fact.metadata(),
        "source": "authoritative_source_table",
        "note": _missing(
            "The source table provides a numeric segment fact without a separate narrative note.",
            source_ref=fact.source_ref,
        ),
        "revenue": _field(
            value,
            source_ref=fact.source_ref,
            core=True,
            unit="$m",
            period=fact.period,
            definition="Revenue for the stated segment dimension, member, and fiscal period.",
        ),
    }
    if canonical_segment_period_type(fact.period_type) == "annual":
        item["annual_revenue"] = dict(item["revenue"])
    return item


def _build_segments(
    segment_rows: Sequence[Mapping[str, Any]],
    history_rows: Sequence[Mapping[str, Any]],
    data_root: Path,
    workbook_path: Path,
    demotions: list[dict[str, Any]],
) -> dict[str, Any]:
    legacy_items = _legacy_bs_segment_items(workbook_path)
    display_orders = {
        (str(item["dimension"]), str(item["member"])): int(item["display_order"])
        for item in legacy_items
    }
    fiscal_periods = {
        _to_iso(row.get("quarter")): (str(row.get("fiscal_label") or ""), row.get("fiscal_year"))
        for row in history_rows
        if _is_present(row.get("quarter"))
    }
    rows = [
        row
        for row in segment_rows
        if _is_present(row.get("segment")) and _is_present(row.get("metric")) and _is_present(row.get("value"))
    ]
    rows.sort(
        key=lambda row: (
            _to_iso(row.get("quarter")),
            str(row.get("period_type")),
            str(row.get("segment")),
            str(row.get("metric")),
            str(row.get("source_doc") or row.get("doc") or row.get("source") or ""),
            str(row.get("value") or ""),
        )
    )
    items: list[dict[str, Any]] = list(legacy_items)
    fixed_display_order = {
        ("geography", "Americas"): 1,
        ("geography", "EMEA"): 2,
        ("geography", "APAC"): 3,
        ("total_company", "Total Company"): 1,
        ("brand", "Hollister"): 1,
        ("brand", "Abercrombie"): 2,
    }
    seen_business_keys = {
        canonical_segment_business_identity(item)
        for item in items
    }
    raw_revenue_keys: set[tuple[str, str, str, str, str]] = set()
    for row in rows:
        metric = str(row.get("metric") or "")
        source_ref = str(row.get("source_doc") or row.get("doc") or _source_ref("Slides_Segments", row, workbook_path=workbook_path))
        source_unit = str(row.get("unit") or "").strip()
        source_scale = "millions" if source_unit.lower() in {"$m", "usdm"} else "ones" if source_unit.lower() in {"usd", "$"} else "not_applicable"
        value = (
            normalize_segment_currency_to_millions(
                row.get("value"),
                source_unit=source_unit,
                source_scale=source_scale,
            )
            if metric == "revenue"
            else row.get("value")
        )
        note = _visible_text_or_blank(
            _clean_text(row.get("note") or row.get("commentary"), limit=220),
            field=f"segments.items.{len(items)}.note",
            section="segments",
            source_ref=source_ref,
            demotions=demotions,
        )
        member = str(row.get("segment") or "").strip()
        period_type = canonical_segment_period_type(row.get("period_type") or "quarterly")
        source_period_end = _to_iso(row.get("quarter"))
        fiscal_label, fiscal_year = fiscal_periods.get(source_period_end, ("", None))
        normalized_period = f"{fiscal_year}-FY" if period_type == "annual" and fiscal_year not in (None, "") else fiscal_label
        if not normalized_period:
            normalized_period = _normalize_period(row.get("quarter"), period_type=period_type)
        dimension = _segment_dimension(member)
        source_row_ref = _source_ref("Slides_Segments", row, workbook_path=workbook_path)
        business_key = canonical_segment_business_identity(
            {
                "dimension": dimension,
                "member": member,
                "period": normalized_period,
                "period_type": period_type,
                "metric": metric,
            }
        )
        if metric == "revenue" and business_key in raw_revenue_keys:
            raise SegmentNormalizationError(
                f"Duplicate canonical segment revenue identity {business_key!r}; source row {source_row_ref}."
            )
        if metric == "revenue":
            raw_revenue_keys.add(business_key)
        if business_key in seen_business_keys:
            continue
        item = {
            "dimension": dimension,
            "member": member,
            "display_order": display_orders.get((_segment_dimension(member), member), 999),
            "segment": _field(member, source_ref=source_ref, core=True),
            "metric": metric,
            "period": normalized_period,
            "period_type": period_type,
            "unit": "$m" if metric == "revenue" else source_unit,
            "source_unit": source_unit,
            "source_scale": source_scale,
            "source_table_scope": period_type,
            "source_table_id": f"{Path(source_ref).name}:{str(row.get('source_type') or row.get('source') or 'legacy_adapter')}",
            "source_row_ref": source_row_ref,
            "source_ref": source_ref,
            "aggregation_role": segment_aggregation_role(dimension, member),
            "source": str(row.get("source") or row.get("source_type") or ""),
            "note": _field(note, source_ref=source_ref) if note else _missing("No concise source-backed segment note survived visible-text quality filtering.", source_ref=source_ref),
        }
        if metric == "revenue" and period_type == "annual":
            item["annual_revenue"] = _field(value, source_ref=source_ref, core=True, unit="$m", period=normalized_period)
            item["revenue"] = _field(value, source_ref=source_ref, unit="$m", period=normalized_period)
        elif metric == "revenue":
            item["revenue"] = _field(value, source_ref=source_ref, core=True, unit="$m", period=normalized_period)
        elif "margin" in metric:
            item["margin"] = _field(value, source_ref=source_ref, unit=str(row.get("unit") or ""))
        elif "operating" in metric:
            item["operating_income"] = _field(value, source_ref=source_ref, unit="$m")
        else:
            item["metric_value"] = _field(value, source_ref=source_ref, unit=str(row.get("unit") or ""))
        items.append(item)
        seen_business_keys.add(business_key)

    item_positions = {
        canonical_segment_business_identity(item): index
        for index, item in enumerate(items)
    }
    for fact in _anf_authoritative_segment_source_facts(data_root):
        item = _normalized_segment_revenue_item(
            fact,
            display_order=fixed_display_order[(fact.dimension, fact.member)],
        )
        position = item_positions.get(fact.business_identity)
        if position is None:
            position = len(items)
            item_positions[fact.business_identity] = position
            items.append(item)
        else:
            items[position] = item
    items.sort(key=lambda item: (str(item.get("period_type")), str(item.get("period")), int(item.get("display_order") or 999), str(item.get("dimension")), str(item.get("member"))))
    return {"items": items}


def _build_operating_drivers(
    driver_rows: Sequence[Mapping[str, Any]],
    workbook_path: Path,
    demotions: list[dict[str, Any]],
) -> dict[str, Any]:
    rows = [row for row in driver_rows if _is_present(row.get("Driver")) or _is_present(row.get("Commentary"))]
    rows.sort(key=lambda row: (_to_iso(row.get("Quarter")), str(row.get("Driver group")), str(row.get("Driver"))))
    items: list[dict[str, Any]] = []
    for row in rows:
        source_ref = str(row.get("Source") or _source_ref("operating_drivers_raw", row, workbook_path=workbook_path))
        driver = str(row.get("Driver") or row.get("Driver group") or "").strip()
        commentary = _clean_text(row.get("Commentary") or row.get("Value"), limit=260)
        field_prefix = f"operating_drivers.items.{len(items)}"
        if not _visible_text_is_clean(
            commentary,
            field=f"{field_prefix}.current_read",
            section="operating_drivers",
            source_ref=source_ref,
            demotions=demotions,
        ):
            continue
        items.append(
            {
                "topic": _field(str(row.get("Driver group") or ""), source_ref=source_ref),
                "driver": _field(driver, source_ref=source_ref, core=True),
                "current_read": _field(commentary or driver, source_ref=source_ref, core=True),
                "metric_value": _field(row.get("Value"), source_ref=source_ref, unit=str(row.get("Unit") or "")) if _is_present(row.get("Value")) else _missing("Driver row has commentary but no clean metric value.", source_ref=source_ref),
                "source": source_ref,
                "why_it_matters": _field(_clean_text(row.get("Commentary"), limit=220), source_ref=source_ref),
                "quality": str(row.get("Quality") or ""),
                "period": _normalize_period(row.get("Quarter"), period_type="quarterly"),
                "driver_type": _driver_type(str(row.get("Driver group") or ""), driver),
                "evidence_key": _evidence_key(source_ref, row.get("Driver group"), driver, commentary),
                "display_role": "history",
                "display_priority": 999,
            }
        )
    selected_topics: set[str] = set()
    priority = 0
    for item in sorted(items, key=lambda value: (str(value.get("period") or ""), str(value.get("evidence_key") or "")), reverse=True):
        topic = str(item.get("topic", {}).get("value") or "").strip().lower()
        if not topic or topic in selected_topics or priority >= 4:
            continue
        selected_topics.add(topic)
        priority += 1
        item["display_role"] = "current_watchlist"
        item["display_priority"] = priority
    return {"items": items}


def _build_quarter_notes(
    note_rows: Sequence[Mapping[str, Any]],
    latest_source_period: str,
    workbook_path: Path,
    demotions: list[dict[str, Any]],
) -> dict[str, Any]:
    rows = [row for row in note_rows if _is_present(row.get("note")) or _is_present(row.get("renderable_note"))]
    rows.sort(key=lambda row: (_to_iso(row.get("quarter")), int(row.get("rank") or 999), str(row.get("topic") or "")))
    items: list[dict[str, Any]] = []
    for row in rows:
        source_ref = str(row.get("source_doc") or row.get("doc") or row.get("evidence_doc") or _source_ref("Quarter_Notes", row, workbook_path=workbook_path))
        note = _clean_text(row.get("renderable_note") or row.get("note") or row.get("body") or row.get("claim"), limit=300)
        field_prefix = f"quarter_notes.items.{len(items)}"
        if not _visible_text_is_clean(
            note,
            field=f"{field_prefix}.note",
            section="quarter_notes",
            source_ref=source_ref,
            demotions=demotions,
        ):
            continue
        implication = _clean_text(row.get("render_summary") or row.get("render_change"), limit=220)
        if not _visible_text_is_clean(
            implication,
            field=f"{field_prefix}.model_implication",
            section="quarter_notes",
            source_ref=source_ref,
            demotions=demotions,
        ):
            continue
        quarter = _normalize_period(row.get("quarter") or row.get("quarter_end"), period_type="quarterly")
        display_role = "audit_only" if latest_source_period and quarter > latest_source_period else "history"
        if display_role == "audit_only":
            demotions.append(
                {
                    "severity": "P2",
                    "rule_id": "quarter_note_after_available_evidence",
                    "field": f"{field_prefix}.quarter",
                    "message": f"Quarter note period {quarter} is later than available financial evidence {latest_source_period}.",
                    "source_ref": source_ref,
                    "suggested_action": "Keep the row audit-only until period evidence is available.",
                }
            )
        items.append(
            {
                "theme": _field(str(row.get("topic") or row.get("tag") or ""), source_ref=source_ref),
                "quarter": _field(quarter, source_ref=source_ref),
                "metric": _field(str(row.get("metric_ref") or row.get("category") or ""), source_ref=source_ref),
                "note": _field(note, source_ref=source_ref, core=True),
                "commentary": _field(note, source_ref=source_ref),
                "model_implication": _field(implication, source_ref=source_ref),
                "valuation_implication": _field(_clean_text(row.get("render_bucket"), limit=160), source_ref=source_ref) if _clean_text(row.get("render_bucket"), limit=160) else _missing("No source-backed valuation implication was normalized.", source_ref=source_ref),
                "source": source_ref,
                "confidence": str(row.get("confidence") or ""),
                "evidence_key": _evidence_key(source_ref, row.get("quarter") or row.get("quarter_end"), row.get("topic") or row.get("tag"), row.get("metric_ref") or row.get("category"), note),
                "display_role": display_role,
                "display_priority": 999,
                "selection_rank": int(row.get("rank") or 999),
            }
        )
    valid_periods = [str(item["quarter"].get("value") or "") for item in items if item["display_role"] != "audit_only"]
    if valid_periods:
        latest_period = max(valid_periods)
        seen_themes: set[str] = set()
        priority = 0
        for item in sorted(items, key=lambda value: (int(value.get("selection_rank") or 999), str(value.get("evidence_key") or ""))):
            if str(item["quarter"].get("value") or "") != latest_period or item["display_role"] == "audit_only":
                continue
            theme = re.sub(r"[^a-z0-9]+", "", str(item["theme"].get("value") or "").lower())
            if not theme or theme in seen_themes or priority >= 6:
                continue
            seen_themes.add(theme)
            priority += 1
            item["display_role"] = "current_note"
            item["display_priority"] = priority
    return {"items": items}


def _build_anf_company_profile(workbook_path: Path, latest_annual_period: str) -> dict[str, Any]:
    description_refs = (_anf_annual_report_ref(6),)
    strategy_refs = (_anf_transcript_ref(30, 32), _anf_transcript_ref(52, 62))
    advantage_refs = (_anf_transcript_ref(26, 30), _anf_transcript_ref(313, 319))
    operating_model_rows = [
        {
            "member": "Americas",
            "description": _narrative_field(
                "Fiscal 2025 sales grew 7%, supported by cross-channel traffic, marketing and store expansion.",
                (_anf_transcript_ref(20, 20),),
                classification="source_backed_fact",
            ),
            "display_order": 1,
        },
        {
            "member": "EMEA",
            "description": _narrative_field(
                "Fiscal 2025 sales grew 6%, led by double-digit growth in the U.K. and growth in the Middle East.",
                (_anf_transcript_ref(20, 20),),
                classification="source_backed_fact",
            ),
            "display_order": 2,
        },
        {
            "member": "APAC",
            "description": _narrative_field(
                "Fiscal 2025 sales grew 5%; management is reviewing strategic alternatives because returns have not fully reflected investment.",
                (_anf_transcript_ref(20, 20), _anf_transcript_ref(52, 52)),
                classification="evidence_backed_synthesis",
            ),
            "display_order": 3,
        },
    ]
    dependency_specs = (
        (
            "Product execution and resonant marketing must sustain traffic and convert demand across both brands.",
            (_anf_transcript_ref(313, 319),),
        ),
        (
            "Tariff mitigation, freight, pricing and sourcing must support the 12.0%-12.5% operating-margin guide.",
            (_anf_transcript_ref(54, 60), _anf_transcript_ref(290, 292)),
        ),
        (
            "The merchandising ERP implementation must remain temporary and avoid lasting sales or inventory disruption.",
            (_anf_transcript_ref(58, 60),),
        ),
        (
            "The read-and-react inventory model must preserve chase capacity and healthy AUR without excess markdown risk.",
            (_anf_transcript_ref(28, 28), _anf_transcript_ref(303, 305)),
        ),
        (
            "Free cash flow and liquidity must support repurchases without weakening financial flexibility.",
            (_anf_transcript_ref(48, 48), _anf_transcript_ref(56, 58)),
        ),
    )
    key_dependencies = [
        {
            "business_key": f"dependency-{index}",
            "text": _narrative_field(text, refs, classification="evidence_backed_synthesis"),
            "display_order": index,
        }
        for index, (text, refs) in enumerate(dependency_specs, start=1)
    ]
    source_ref = f"{workbook_path.name}!SUMMARY"
    return {
        "company_name": _field("Abercrombie & Fitch Co.", source_ref=source_ref, core=True),
        "sector": _field("Consumer Discretionary", source_ref=source_ref, core=True),
        "industry": _field("Specialty apparel retail", source_ref=source_ref, core=True),
        "business_description": _narrative_field(
            "Abercrombie & Fitch is a global omnichannel specialty retailer operating the Abercrombie and Hollister brand families through stores, digital channels and third-party partnerships.",
            description_refs,
            classification="source_backed_fact",
            core=True,
        ),
        "strategic_context": _narrative_field(
            "ANF enters 2026 from record 2025 sales with both brands expected to grow, while management works through tariff pressure, an ERP launch and a lower operating-margin guide.",
            strategy_refs,
            classification="evidence_backed_synthesis",
            core=True,
        ),
        "revenue_model": _narrative_field(
            "Revenue is generated across two brand families, owned stores, digital channels, geographic regions and growing third-party partnerships.",
            (_anf_transcript_ref(26, 30), _anf_transcript_ref(52, 52)),
            classification="evidence_backed_synthesis",
            core=True,
        ),
        "revenue_mix_label": _field(
            "Revenue mix by geography (% of revenue)",
            source_ref=f"{workbook_path.name}!SUMMARY!A8",
            core=True,
        ),
        "revenue_streams": _build_revenue_stream_rows(workbook_path, period=latest_annual_period),
        "key_advantages": _narrative_field(
            "Two scaled brands, profitable stores and digital channels, direct customer reach and a read-and-react inventory model support agile demand response.",
            advantage_refs,
            classification="evidence_backed_synthesis",
        ),
        "key_risks": _narrative_field(
            "Execution risk centers on fashion demand, brand momentum, tariffs, inventory quality, ERP disruption and disciplined capital allocation.",
            (_anf_transcript_ref(28, 32), _anf_transcript_ref(54, 62)),
            classification="evidence_backed_synthesis",
        ),
        "operating_model_rows": operating_model_rows,
        "key_dependencies": key_dependencies,
        "allowed_sector_terms": ["Abercrombie", "Hollister", "APAC", "EMEA", "Americas"],
    }


def _scenario_source_refs(source_ref: str, evidence_refs: Sequence[Any] = ()) -> list[str]:
    refs = [str(value) for value in evidence_refs if str(value)]
    if source_ref and source_ref not in refs:
        refs.insert(0, source_ref)
    return refs or ["unavailable_source_reference"]


def _build_typed_scenario_contract(
    valuation_inputs: Mapping[str, Any],
    guidance_items: Sequence[Mapping[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    as_of_field = valuation_inputs.get("as_of_date") if isinstance(valuation_inputs.get("as_of_date"), Mapping) else {}
    authoritative_as_of_date = (
        str(as_of_field.get("value") or "")
        if str(as_of_field.get("status") or "") == "populated"
        else None
    )
    actual_specs = (
        ("price", "price", "Current share price", "$/share", "as_of", 0.0, None),
        ("shares_outstanding", "shares_outstanding", "Shares outstanding", "m shares", "as_of", 0.0, None),
        ("diluted_shares", "diluted_shares", "Diluted shares", "m shares", "reported_period", 0.0, None),
        ("net_debt", "net_debt", "Net debt", "$m", "as_of", None, None),
        ("revenue_ttm", "revenue_ttm", "Revenue TTM", "$m", "TTM", 0.0, None),
        ("base_ebitda_ttm", "base_ebitda_ttm", "EBITDA (base, TTM)", "$m", "TTM", None, None),
        ("adjusted_ebitda_ttm", "adjusted_ebitda_ttm", "Adjusted EBITDA TTM", "$m", "TTM", None, None),
        ("free_cash_flow_ttm", "fcf_ttm", "Free cash flow TTM", "$m", "TTM", None, None),
        ("net_income_ttm", "net_income_ttm", "Net income TTM", "$m", "TTM", None, None),
    )
    items: list[dict[str, Any]] = []
    for field_name, assumption_id, metric, unit, default_horizon, minimum, maximum in actual_specs:
        field = valuation_inputs.get(field_name) if isinstance(valuation_inputs.get(field_name), Mapping) else {}
        status = str(field.get("status") or "missing_source")
        value = field.get("value") if status == "populated" else None
        source_ref = str(field.get("source_ref") or f"valuation_inputs.{field_name}")
        horizon = (
            str(field.get("period") or "")
            if default_horizon == "reported_period"
            else default_horizon
        )
        if not horizon:
            horizon = "TTM" if default_horizon == "reported_period" else default_horizon
        items.append(
            {
                "scenario_id": "common",
                "assumption_id": assumption_id,
                "metric": metric,
                "value_kind": "point" if status == "populated" else "unavailable",
                "value": value,
                "low_value": None,
                "high_value": None,
                "unit": str(field.get("unit") or unit),
                "horizon": horizon,
                "dimension_id": "total_company",
                "member": "total_company",
                "profile_pack_id": None,
                "as_of_date": authoritative_as_of_date if horizon == "as_of" else None,
                "source_classification": "source_actual" if status == "populated" else "unavailable",
                "validation": {"minimum": minimum, "maximum": maximum},
                "propagation_rule": "shared_actual" if status == "populated" else "no_propagation",
                "status": status,
                "source_ref": source_ref,
                "source_refs": _scenario_source_refs(source_ref, field.get("evidence_refs") or ()),
                "reason": str(field.get("reason") or ""),
            }
        )

    denominator = valuation_inputs.get("per_share_denominator") if isinstance(valuation_inputs.get("per_share_denominator"), Mapping) else {}
    denominator_ref = str(denominator.get("source_ref") or "user_input")
    items.append(
        {
            "scenario_id": "common",
            "assumption_id": "per_share_mode",
            "metric": "Per-share denominator mode",
            "value_kind": "unavailable",
            "value": None,
            "low_value": None,
            "high_value": None,
            "unit": "classification",
            "horizon": "as_of",
            "dimension_id": "total_company",
            "member": "total_company",
            "profile_pack_id": None,
            "as_of_date": None,
            "source_classification": "user_input",
            "validation": {"minimum": None, "maximum": None},
            "propagation_rule": "scenario_specific",
            "status": "missing_source",
            "source_ref": denominator_ref,
            "source_refs": [denominator_ref],
            "reason": "Outstanding or Diluted must be selected explicitly; no denominator fallback is permitted.",
        }
    )

    guidance_specs = {
        "adjusted_eps": ("adjusted_eps_guidance", "Adjusted diluted EPS guidance", "adjusted_eps", "manual_incremental", "direct_eps", "non_cash"),
        "capital_expenditures": ("capital_expenditures_guidance", "Capital expenditures guidance", "capital_expenditures", "cash_flow_capex", "cash_only", "investing"),
        "operating_margin": ("operating_margin_guidance", "Operating margin guidance", "operating_margin", "margin_ebitda", "manual_review_required", "manual_review_required"),
        "revenue": ("revenue_growth", "Revenue growth guidance", "revenue", "revenue_volume", "manual_review_required", "operating"),
    }
    bridges: list[dict[str, Any]] = []
    seen_guidance: set[tuple[str, str, str]] = set()
    for row in guidance_items:
        if str(row.get("display_role") or "") not in {"current_primary", "current_secondary"}:
            continue
        comparison = row.get("comparison_contract")
        if not isinstance(comparison, Mapping):
            continue
        metric_id = str(comparison.get("metric") or "")
        spec = guidance_specs.get(metric_id)
        if spec is None:
            continue
        assumption_id, display_metric, impact_metric, driver_type, tax_treatment, cash_classification = spec
        horizon = str(comparison.get("horizon") or "")
        key = (assumption_id, horizon, str(row.get("publication_date") or ""))
        if key in seen_guidance:
            continue
        seen_guidance.add(key)
        comparison_type = str(comparison.get("comparison_type") or "")
        unit = str(comparison.get("unit") or "")
        scale = 0.01 if unit == "%" else 1.0
        point = comparison.get("value")
        low = comparison.get("low")
        high = comparison.get("high")
        value_kind = comparison_type if comparison_type in {"range", "minimum", "maximum"} else "point"
        value = float(point) * scale if isinstance(point, (int, float)) else None
        low_value = float(low) * scale if isinstance(low, (int, float)) else None
        high_value = float(high) * scale if isinstance(high, (int, float)) else None
        source_ref = str(comparison.get("source_ref") or row.get("source_ref") or "")
        source_refs = _scenario_source_refs(source_ref, row.get("evidence_refs") or ())
        items.append(
            {
                "scenario_id": "common",
                "assumption_id": assumption_id,
                "metric": display_metric,
                "value_kind": value_kind,
                "value": value,
                "low_value": low_value,
                "high_value": high_value,
                "unit": unit,
                "horizon": horizon,
                "dimension_id": "total_company",
                "member": "total_company",
                "profile_pack_id": None,
                "as_of_date": None,
                "source_classification": "source_guidance",
                "validation": {"minimum": None, "maximum": None},
                "propagation_rule": "no_propagation",
                "status": "populated",
                "source_ref": source_ref,
                "source_refs": source_refs,
                "publication_date": row.get("publication_date"),
                "reporting_period": row.get("stated_in_period"),
                "reason": "A source range or threshold is preserved but not converted into an unsupported scenario point.",
            }
        )
        bridge_value = value if value_kind in {"point", "minimum", "maximum"} else None
        bridges.append(
            {
                "scenario_id": "common",
                "driver_id": assumption_id,
                "profile_pack_id": "retail_operating_pack",
                "driver_type": driver_type,
                "metric": display_metric,
                "impact_metric": impact_metric,
                "value": bridge_value,
                "unit": unit,
                "horizon": horizon,
                "dimension_id": "total_company",
                "member": "total_company",
                "source_classification": "source_guidance",
                "tax_treatment": tax_treatment,
                "cash_classification": cash_classification,
                "propagation_rule": "no_propagation",
                "status": "populated" if bridge_value is not None else "manual_review_required",
                "source_ref": source_ref,
                "source_refs": source_refs,
                "reason": "Range guidance requires an explicit user-selected scenario point before it can affect formulas." if bridge_value is None else "Typed threshold is retained for review and is not auto-propagated.",
            }
        )
    canonical, token_issues = canonicalize_scenario_contract(
        {"scenario_items": items, "scenario_driver_bridge": bridges},
        allowed_profile_pack_ids={"retail_operating_pack"},
        allowed_scenario_driver_ids={str(row.get("driver_id") or "") for row in bridges},
        allowed_dimension_ids={"total_company"},
    )
    if token_issues:
        raise ValueError(
            "ANF scenario token normalization failed: "
            + "; ".join(f"{issue.field}: {issue.message}" for issue in token_issues)
        )
    return canonical["scenario_items"], canonical["scenario_driver_bridge"]


def _build_investment_case(
    valuation_inputs: Mapping[str, Any],
    guidance_items: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    model_refs = (_anf_transcript_ref(14, 32), _anf_transcript_ref(52, 62), _anf_transcript_ref(374, 378))
    summary = _narrative_field(
        "ANF combines balanced brand and regional growth with strong cash generation; the investment case now depends on proving that double-digit margins can persist through 2026 cost pressure.",
        model_refs,
        classification="evidence_backed_synthesis",
        core=True,
    )
    key_debate = _narrative_field(
        "Can ANF sustain structurally higher margins and earnings while absorbing tariffs, ERP disruption and tougher comparisons?",
        (_anf_transcript_ref(54, 62), _anf_transcript_ref(370, 378)),
        classification="analyst_interpretation_requiring_review",
        core=True,
        review_state="manual_review_required",
    )
    invalidators = [
        {
            "business_key": "sales-execution-breaks",
            "text": _narrative_field(
                "The thesis weakens if product and marketing execution fail to keep sales within the 2026 growth range across both brands.",
                (_anf_transcript_ref(52, 52), _anf_transcript_ref(313, 319)),
                classification="analyst_interpretation_requiring_review",
                review_state="manual_review_required",
            ),
            "display_order": 1,
        },
        {
            "business_key": "margin-durability-breaks",
            "text": _narrative_field(
                "The thesis weakens if tariff and ERP mitigation fail and operating margin falls materially below the 2026 range.",
                (_anf_transcript_ref(54, 60), _anf_transcript_ref(290, 292)),
                classification="analyst_interpretation_requiring_review",
                review_state="manual_review_required",
            ),
            "display_order": 2,
        },
    ]
    scenario_items, scenario_driver_bridge = _build_typed_scenario_contract(valuation_inputs, guidance_items)
    return {
        "summary": summary,
        "why_it_can_work": _narrative_field(
            "Both brands ended Q4 at record sales, digital reached 44% of annual sales and management describes stores and digital as highly profitable.",
            (_anf_transcript_ref(14, 26), _anf_transcript_ref(50, 50)),
            classification="source_backed_fact",
        ),
        "key_debate": key_debate,
        "upside_factors": _narrative_field(
            "Upside requires product execution to keep both brands growing while freight, pricing and sourcing offset more of the tariff burden than assumed.",
            (_anf_transcript_ref(290, 292), _anf_transcript_ref(313, 319)),
            classification="analyst_interpretation_requiring_review",
            review_state="manual_review_required",
        ),
        "downside_factors": _narrative_field(
            "Downside centers on weaker demand, persistent tariff costs, ERP disruption or inventory pressure that pushes margins below guidance.",
            (_anf_transcript_ref(54, 62), _anf_transcript_ref(303, 305)),
            classification="analyst_interpretation_requiring_review",
            review_state="manual_review_required",
        ),
        "watch_next": _narrative_field(
            "Watch Q1 brand growth, the tariff-freight-ERP margin bridge, inventory units after the prebuild and free-cash-flow coverage of buybacks.",
            (_anf_transcript_ref(58, 62), _anf_transcript_ref(303, 305), _anf_transcript_ref(48, 48)),
            classification="analyst_interpretation_requiring_review",
            review_state="manual_review_required",
        ),
        "current_stance": _narrative_field(
            "Constructive on the operating model and cash generation, but margin-sensitive until 2026 mitigation is demonstrated.",
            model_refs,
            classification="analyst_interpretation_requiring_review",
            review_state="manual_review_required",
        ),
        "bull_case": _missing("No source-backed bull-case valuation assumption is available.", source_ref=model_refs[0]),
        "base_case": _missing("No source-backed base-case valuation assumption is available.", source_ref=model_refs[0]),
        "bear_case": _missing("No source-backed bear-case valuation assumption is available.", source_ref=model_refs[0]),
        "scenario_items": scenario_items,
        "scenario_driver_bridge": scenario_driver_bridge,
        "invalidators": invalidators,
        "source_evidence": [
            {"source_ref": ref, "section": "Q4 2025 earnings call"}
            for ref in model_refs
        ],
    }


def _legacy_promise_scope(row: Mapping[str, Any]) -> tuple[str, int] | None:
    metric_text = str(row.get("metric_display") or row.get("metric_ref") or "").strip()
    metric_text = re.sub(r"\s+guidance$", "", metric_text, flags=re.I)
    metric_text = {
        "eps": "Adj EPS",
        "diluted-share": "Diluted shares",
        "real-estate": "Real estate activity",
        "tariff-impact": "Tariffs",
    }.get(metric_text.casefold(), metric_text)
    horizon = str(row.get("target_period_label") or row.get("target_period_norm") or "").strip()
    scope = normalize_guidance_scope(
        {
            "metric": {"value": metric_text},
            "horizon": {"value": horizon},
            "value": {"unit": ""},
        }
    )
    if not scope.metric or scope.fiscal_year is None or scope.horizon_type != "FY":
        return None
    return scope.metric, int(scope.fiscal_year)


def _legacy_promise_evidence_dispositions(
    promise_rows: Sequence[Mapping[str, Any]],
    *,
    represented_keys: set[tuple[str, int]],
    workbook_path: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Preserve every source-evidenced legacy Promise occurrence outside active routes."""

    dispositions: list[dict[str, Any]] = []
    first_by_signature: dict[tuple[tuple[str, int], str, str, str], str] = {}
    for row in promise_rows:
        scope_key = _legacy_promise_scope(row)
        if scope_key is None or scope_key in represented_keys:
            continue
        metric, fiscal_year = scope_key
        try:
            evidence = json.loads(str(row.get("source_evidence_json") or "{}"))
        except json.JSONDecodeError:
            evidence = {}
        if not isinstance(evidence, Mapping):
            evidence = {}

        legacy_ref = _source_ref("Promise_Progress", row, workbook_path=workbook_path)
        source_document = str(evidence.get("doc") or "").strip()
        source_ref = f"{source_document}#{legacy_ref}" if source_document else legacy_ref
        source_excerpt = _clean_text(evidence.get("line"), limit=500)
        target_value = str(row.get("target_display") or row.get("target") or "").strip()
        promise_id = str(row.get("promise_id") or f"{metric}:FY{fiscal_year}:row:{row.get('_row_number')}")
        business_key = f"{metric}:FY{fiscal_year}"
        source_exists = bool(source_document and Path(source_document).is_file())
        signature = (scope_key, source_document, source_excerpt, target_value)
        related_promise_id = ""

        if not source_exists or not source_excerpt:
            disposition = "unavailable_without_adequate_evidence"
            reason = "The legacy Promise row lacks a resolvable source document or source excerpt for the claimed business meaning."
            supports_claim = False
            review_state = "manual_review_required"
        elif scope_key in _LEGACY_PROMISE_REJECTION_REASONS:
            disposition = "rejected_with_evidence"
            reason = _LEGACY_PROMISE_REJECTION_REASONS[scope_key]
            supports_claim = False
            review_state = "rejected"
        elif scope_key == ("operating_margin", 2023):
            disposition = "duplicate_or_superseded_evidence"
            reason = (
                "The source identifies 8%-9% as the previous fiscal-2023 operating-margin outlook and replaces it "
                "with around 10%; the prior range is retained as superseded evidence only."
            )
            supports_claim = True
            review_state = "accepted"
        elif scope_key in _LEGACY_PROMISE_AUDIT_ONLY_KEYS:
            related_promise_id = first_by_signature.get(signature, "")
            if related_promise_id:
                disposition = "duplicate_or_superseded_evidence"
                reason = "The same source document, excerpt and historical Promise value are already retained by another occurrence."
            else:
                disposition = "audit_only_historical_evidence"
                reason = (
                    "The source supports this historical Promise value, but it predates the visible progression blocks "
                    "and is retained in JSON audit history rather than duplicated in the workbook."
                )
                first_by_signature[signature] = promise_id
            supports_claim = True
            review_state = "accepted"
        else:
            disposition = "rejected_with_evidence"
            reason = (
                "Source evidence exists, but no definition-compatible metric and horizon route has been established; "
                "the occurrence is retained and rejected from visible Promise content."
            )
            supports_claim = False
            review_state = "rejected"

        dispositions.append(
            {
                "business_key": business_key,
                "promise_id": promise_id,
                "metric": metric,
                "horizon": f"FY{fiscal_year}",
                "legacy_evidence_date": _to_iso(
                    row.get("last_seen_evidence_quarter")
                    or row.get("first_seen_evidence_quarter")
                    or row.get("quarter")
                ),
                "target_value": target_value,
                "source_ref": source_ref,
                "source_refs": [source_ref],
                "source_document": source_document,
                "source_excerpt": source_excerpt,
                "source_evidence": dict(evidence),
                "legacy_row_number": int(row.get("_row_number") or 0),
                "disposition": disposition,
                "disposition_reason": reason,
                "related_promise_id": related_promise_id,
                "supports_claimed_business_meaning": supports_claim,
                "visibility_disposition": (
                    "audit_only"
                    if disposition in {"audit_only_historical_evidence", "duplicate_or_superseded_evidence"}
                    else "rejected"
                    if disposition == "rejected_with_evidence"
                    else "unavailable"
                ),
                "review_state": review_state,
                "evidence_key": _evidence_key("legacy_promise_disposition", promise_id, source_ref, source_excerpt),
            }
        )

    counts: dict[str, int] = defaultdict(int)
    for row in dispositions:
        counts[str(row["disposition"])] += 1
    summary = {
        "business_key_count": len({row["business_key"] for row in dispositions}),
        "occurrence_count": len(dispositions),
        "disposition_counts": dict(sorted(counts.items())),
    }
    return dispositions, summary


def _progress_actual(
    *,
    fiscal_year: int,
    metric: str,
    annual_rows: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    result: dict[str, Any]
    if fiscal_year == 2025 and metric == "Revenue":
        result = _field(6.0, source_ref=_anf_transcript_ref(16, 18), unit="%", period="2025-FY", evidence_refs=(_anf_transcript_ref(16, 18),))
    elif fiscal_year == 2025 and metric == "Adj EPS":
        result = _field(9.86, source_ref=_anf_transcript_ref(48, 48), unit="$/share", period="2025-FY", evidence_refs=(_anf_transcript_ref(48, 48),))
    elif fiscal_year == 2025 and metric == "Real estate activity":
        result = _field(40.0, source_ref=_anf_transcript_ref(50, 50), unit="stores", period="2025-FY", evidence_refs=(_anf_transcript_ref(50, 50),), definition="Net store openings derived as 62 openings less 22 closures.")
    else:
        annual = next((row for row in annual_rows if str(row.get("period") or "") == f"{fiscal_year}-FY"), None)
        field_name = {"Capex": "capital_expenditures", "Share repurchases": "buybacks_cash"}.get(metric)
        if annual is not None and field_name and isinstance(annual.get(field_name), Mapping):
            field = annual[field_name]
            if field.get("status") == "populated" and field.get("value") not in (None, ""):
                result = dict(field)
            else:
                result = _missing(
                    "No definition-compatible source-backed annual actual is available for this guidance progression row.",
                    source_ref=f"promise_progress:{fiscal_year}:{metric}",
                )
        else:
            result = _missing(
                "No definition-compatible source-backed annual actual is available for this guidance progression row.",
                source_ref=f"promise_progress:{fiscal_year}:{metric}",
            )
    result["comparison_metric"] = normalize_guidance_scope(
        {"metric": {"value": metric}, "horizon": {"value": f"FY{fiscal_year}"}, "value": {"unit": str(result.get("unit") or "")}}
    ).metric
    result["comparison_horizon"] = f"FY{fiscal_year}"
    return result


def _progress_status(guidance_item: Mapping[str, Any], actual: Mapping[str, Any]) -> tuple[dict[str, Any], str]:
    current_guide = guidance_item.get("value") if isinstance(guidance_item.get("value"), Mapping) else {}
    source_ref = str(current_guide.get("source_ref") or "")
    evidence_refs = tuple(guidance_item.get("evidence_refs") or current_guide.get("evidence_refs") or ())
    if actual.get("status") != "populated" or actual.get("value") in (None, ""):
        return (
            _field(None, status="manual_review_required", source_ref=source_ref, reason="Status requires a definition-compatible actual value.", confidence=""),
            "",
        )
    contract = guidance_item.get("comparison_contract")
    if isinstance(contract, Mapping) and contract.get("comparison_type") == "range":
        low, high = contract.get("low"), contract.get("high")
        actual_value = actual.get("value")
        contract_scope = normalize_guidance_scope(
            {
                "metric": {"value": contract.get("metric")},
                "horizon": {"value": contract.get("horizon")},
                "value": {"unit": contract.get("unit")},
            }
        )
        actual_scope = normalize_guidance_scope(
            {
                "metric": {"value": actual.get("comparison_metric")},
                "horizon": {"value": actual.get("comparison_horizon")},
                "value": {"unit": actual.get("unit")},
            }
        )
        compatible = (
            isinstance(low, (int, float))
            and not isinstance(low, bool)
            and isinstance(high, (int, float))
            and not isinstance(high, bool)
            and isinstance(actual_value, (int, float))
            and not isinstance(actual_value, bool)
            and float(low) <= float(high)
            and contract_scope.metric == actual_scope.metric
            and contract_scope.horizon == actual_scope.horizon
            and bool(contract_scope.unit)
            and contract_scope.unit == actual_scope.unit
        )
        if compatible:
            value = "Within range" if float(low) <= float(actual_value) <= float(high) else "Outside range"
            return (
                _field(value, source_ref=source_ref, evidence_refs=evidence_refs),
                "actual_within_published_range",
            )
    return (
        _field(
            None,
            status="manual_review_required",
            source_ref=source_ref,
            reason="Status requires an explicit same-metric, same-unit, same-horizon typed comparison contract.",
            confidence="",
            evidence_refs=evidence_refs,
        ),
        "",
    )


def _build_promise_progress(
    guidance_items: Sequence[Mapping[str, Any]],
    annual_rows: Sequence[Mapping[str, Any]],
    promise_rows: Sequence[Mapping[str, Any]],
    workbook_path: Path,
) -> dict[str, Any]:
    capacities = {2025: 8, 2024: 4, 2023: 3}
    metric_priority = {
        "Revenue": 1,
        "Operating margin": 2,
        "Adj EPS": 3,
        "Share repurchases": 4,
        "Diluted shares": 5,
        "Capex": 6,
        "Real estate activity": 7,
        "Tariffs": 8,
    }
    grouped: dict[tuple[int, str], list[Mapping[str, Any]]] = defaultdict(list)
    for item in guidance_items:
        scope = normalize_guidance_scope(item)
        if scope.horizon_type != "FY" or scope.fiscal_year not in capacities:
            continue
        if str(item.get("display_role") or "") == "audit_only":
            continue
        metric = str((item.get("metric") or {}).get("value") or "")
        grouped[(int(scope.fiscal_year), metric)].append(item)

    rows: list[dict[str, Any]] = []
    for (fiscal_year, metric), updates in sorted(grouped.items()):
        updates = sorted(updates, key=lambda item: (str(item.get("publication_date") or ""), str(item.get("evidence_key") or "")))
        first, latest = updates[0], updates[-1]
        refs = list(dict.fromkeys(ref for item in updates for ref in (item.get("evidence_refs") or []) if ref))
        by_reporting_quarter: dict[str, Mapping[str, Any]] = {}
        for item in updates:
            by_reporting_quarter[str(item.get("stated_in_period") or "")] = item

        def update_field(quarter: int) -> dict[str, Any]:
            item = by_reporting_quarter.get(f"{fiscal_year}-Q{quarter}")
            return dict(item["value"]) if item and isinstance(item.get("value"), Mapping) else _missing(
                f"No source-backed Q{quarter} update was normalized for FY{fiscal_year} {metric}.",
                source_ref=refs[0] if refs else "",
            )

        actual = _progress_actual(fiscal_year=fiscal_year, metric=metric, annual_rows=annual_rows)
        status, status_rule_id = _progress_status(latest, actual)
        priority = metric_priority.get(metric, 999)
        visible = priority <= capacities[fiscal_year]
        rows.append(
            {
                "metric": dict(first["metric"]),
                "display_metric": _field(f"FY{fiscal_year} {metric}", source_ref=refs[0] if refs else "", evidence_refs=refs),
                "original_commitment": dict(first["value"]),
                "prior_update": dict(updates[-2]["value"]) if len(updates) > 1 else _missing("No prior update exists before the latest value.", source_ref=refs[0] if refs else ""),
                "current_guidance": dict(latest["value"]),
                "q1_update": update_field(1),
                "q2_update": update_field(2),
                "q3_update": update_field(3),
                "q4_update": update_field(4),
                "actual": actual,
                "progress_status": status,
                "status_rule_id": status_rule_id,
                "status_comparison": dict(latest.get("comparison_contract") or {}),
                "why_it_matters": _narrative_field(
                    "This row preserves the published guidance path and separates factual progression from any unsupported performance judgment.",
                    refs,
                    classification="evidence_backed_synthesis",
                ),
                "horizon": _field(f"FY{fiscal_year}", source_ref=refs[0] if refs else "", evidence_refs=refs),
                "reporting_period": str(latest.get("stated_in_period") or ""),
                "publication_date": str(latest.get("publication_date") or ""),
                "notes_source": _field(
                    f"{len(updates)} source-backed guidance update(s); status uses only an explicit comparison rule.",
                    source_ref=refs[0] if refs else "",
                    evidence_refs=refs,
                ),
                "evidence_refs": refs,
                "evidence_key": _evidence_key("promise_progress", fiscal_year, metric, *refs),
                "display_block": f"fy{fiscal_year}",
                "display_role": "historical_progression" if visible else "audit_only",
                "display_priority": priority,
                "visibility_disposition": "visible" if visible else "audit_only",
                "disposition_reason": "selected_for_historical_progression" if visible else "historical_progression_outside_visible_priority",
                "review_state": "accepted" if status.get("status") == "populated" else "manual_review_required",
            }
        )
    represented_keys: set[tuple[str, int]] = set()
    for item in guidance_items:
        scope = normalize_guidance_scope(item)
        if scope.fiscal_year is not None and scope.horizon_type == "FY":
            represented_keys.add((scope.metric, int(scope.fiscal_year)))
    historical_evidence, historical_summary = _legacy_promise_evidence_dispositions(
        promise_rows,
        represented_keys=represented_keys,
        workbook_path=workbook_path,
    )
    return {
        "items": rows,
        "historical_evidence_items": historical_evidence,
        "historical_evidence_summary": historical_summary,
        "scorecard_items": [],
        "scorecard_disposition": "No source-backed or deterministic generic credibility score is available; the scorecard remains blank.",
    }


def _collect_legacy_unit_reviews(value: Any, path: str = "$") -> list[dict[str, Any]]:
    reviews: list[dict[str, Any]] = []
    if isinstance(value, Mapping):
        legacy_unit = str(value.get("legacy_unit") or "")
        if legacy_unit:
            reviews.append(
                {
                    "severity": "P2",
                    "rule_id": "legacy_adapter_unit_normalization",
                    "field": path,
                    "message": f"Legacy unit {legacy_unit!r} was normalized to {value.get('unit')!r} for schema compatibility.",
                    "source_ref": str(value.get("source_ref") or ""),
                    "suggested_action": "Confirm the unit taxonomy in a source-native extractor before promotion.",
                }
            )
        for key, child in value.items():
            reviews.extend(_collect_legacy_unit_reviews(child, f"{path}.{key}"))
    elif isinstance(value, list):
        for idx, child in enumerate(value):
            reviews.extend(_collect_legacy_unit_reviews(child, f"{path}.{idx}"))
    return reviews


def build_anf_normalized_package(*, data_root: Path, workbook_path: Path) -> dict[str, Any]:
    """Build ANF shadow data from legacy workbook artifacts for comparison only."""
    history_rows = _read_sheet_rows(workbook_path, "History_Q")
    leverage_rows = _read_sheet_rows(workbook_path, "Leverage_Liquidity")
    revolver_rows = _read_sheet_rows(workbook_path, "Revolver_History")
    guidance_rows = _read_sheet_rows(workbook_path, "Guidance_Normalized")
    promise_rows = _read_sheet_rows(workbook_path, "Promise_Progress")
    segment_rows = _read_sheet_rows(workbook_path, "Slides_Segments")
    driver_rows = _read_sheet_rows(workbook_path, "operating_drivers_raw")
    quarter_note_rows = _read_sheet_rows(workbook_path, "Quarter_Notes")
    text_quality_demotions: list[dict[str, Any]] = []
    guidance_routing_reviews: list[dict[str, Any]] = []
    adapter_review_flags: list[dict[str, Any]] = []
    adapter_truncations: list[dict[str, Any]] = []
    adapter_deduplications: list[dict[str, Any]] = []
    source_coverage = _source_coverage(data_root, workbook_path)
    unsupported_zero_placeholders = _read_legacy_unsupported_zero_placeholders(workbook_path)

    full_quarterly_financials = _build_quarterly_financial_rows(
        history_rows,
        workbook_path,
        unsupported_zero_placeholders=unsupported_zero_placeholders,
        review_flags=adapter_review_flags,
    )
    calculation_history = _build_calculation_history(full_quarterly_financials)
    quarterly_financials = _limit_legacy_adapter_rows(
        full_quarterly_financials,
        limit=12,
        collection_path="quarterly_financials.rows",
        workbook_path=workbook_path,
        truncations=adapter_truncations,
        review_flags=adapter_review_flags,
    )
    incomplete_annual_candidates: list[dict[str, Any]] = []
    annual_financials = _build_annual_financial_rows(
        history_rows,
        workbook_path,
        incomplete_candidates=incomplete_annual_candidates,
        unsupported_zero_placeholders=unsupported_zero_placeholders,
    )
    adapter_review_flags.extend(_annual_missing_component_reviews(annual_financials))
    adapter_review_flags.extend(_annual_incomplete_candidate_reviews(incomplete_annual_candidates))
    debt_liquidity = _build_debt_liquidity(
        history_rows,
        full_quarterly_financials,
        leverage_rows,
        revolver_rows,
        workbook_path,
        adapter_review_flags,
    )
    valuation_inputs = _build_valuation_inputs(
        quarterly_financials,
        debt_liquidity,
        workbook_path,
        review_flags=adapter_review_flags,
    )
    # Guidance remains a complete routed history. Exact duplicate source copies
    # collapse into one semantic row with every source reference retained.
    guidance_items = _build_guidance_items(
        guidance_rows,
        promise_rows,
        workbook_path,
        text_quality_demotions,
        guidance_routing_reviews,
    )
    # The visible legacy segment matrix is itself the migration oracle.  A
    # generic tail limit used here previously discarded annual and older
    # quarterly business keys before the planner could apply its period axes.
    segment_items = _build_segments(
        segment_rows,
        history_rows,
        data_root,
        workbook_path,
        text_quality_demotions,
    )["items"]
    driver_candidates = _dedupe_legacy_adapter_rows(
        _build_operating_drivers(driver_rows, workbook_path, text_quality_demotions)["items"],
        collection_path="operating_drivers.items",
        workbook_path=workbook_path,
        deduplications=adapter_deduplications,
        review_flags=adapter_review_flags,
    )
    driver_items = _limit_legacy_adapter_rows(
        driver_candidates,
        limit=30,
        collection_path="operating_drivers.items",
        workbook_path=workbook_path,
        truncations=adapter_truncations,
        review_flags=adapter_review_flags,
    )
    latest_source_period = str(quarterly_financials[-1].get("period") or "") if quarterly_financials else ""
    for item in driver_items:
        item["display_role"] = "history"
        item["display_priority"] = 999
    curated_drivers = _build_anf_source_backed_operating_drivers(latest_source_period)
    driver_items.extend(curated_drivers["items"])
    quarter_note_candidates = _dedupe_legacy_adapter_rows(
        _build_quarter_notes(
            quarter_note_rows,
            latest_source_period,
            workbook_path,
            text_quality_demotions,
        )["items"],
        collection_path="quarter_notes.items",
        workbook_path=workbook_path,
        deduplications=adapter_deduplications,
        review_flags=adapter_review_flags,
    )
    quarter_note_items = _limit_legacy_adapter_rows(
        quarter_note_candidates,
        limit=40,
        collection_path="quarter_notes.items",
        workbook_path=workbook_path,
        truncations=adapter_truncations,
        review_flags=adapter_review_flags,
    )
    curated_quarter_notes = _build_anf_source_backed_quarter_notes(latest_source_period)
    quarter_note_items.extend(curated_quarter_notes["items"])

    source_ref = f"{workbook_path.name}!SUMMARY"
    latest_annual_period = str(annual_financials[-1].get("period") or "") if annual_financials else ""
    package = {
        "package_version": "0.3.0-anf-shadow",
        "generated_at_utc": _now(),
        "stress_test": True,
        "shadow_package": True,
        "ticker_metadata": {
            "ticker": _field("ANF", source_ref="SEC company_tickers + ANF legacy workbook", core=True),
            "exchange": _field("NYSE", source_ref=source_ref, core=True),
            "cik": _field("0001018840", source_ref="sec_cache/ANF/0001018840", core=True),
            "fiscal_year_end": _field("retail fiscal year ending around late January / early February", source_ref=source_ref, core=True),
            "reporting_currency": _field("USD", source_ref=source_ref, core=True),
            "investment_case_title": _field("ANF Investment Case", source_ref="ticker_metadata.ticker", core=True),
        },
        "company_profile": _build_anf_company_profile(workbook_path, latest_annual_period),
        "quarterly_financials": {"rows": quarterly_financials},
        "calculation_history": calculation_history,
        "annual_financials": {
            "rows": annual_financials,
            "incomplete_candidates": incomplete_annual_candidates,
        },
        "debt_liquidity": debt_liquidity,
        "capital_returns": _build_capital_returns(history_rows, workbook_path),
        "normalized_guidance": {"items": guidance_items},
        "promise_progress": _build_promise_progress(
            guidance_items,
            annual_financials,
            promise_rows,
            workbook_path,
        ),
        "segments": {"items": segment_items},
        "operating_drivers": {"items": driver_items, "current_outlook": curated_drivers["current_outlook"]},
        "quarter_notes": {"items": quarter_note_items, "summary": curated_quarter_notes["summary"]},
        "investment_case": _build_investment_case(valuation_inputs, guidance_items),
        "valuation_inputs": valuation_inputs,
        "valuation_outputs": {"items": []},
        "source_coverage": source_coverage,
        "mapping_gaps": [],
        "manual_review_flags": list(text_quality_demotions) + guidance_routing_reviews + adapter_review_flags,
    }
    unit_reviews = _collect_legacy_unit_reviews(package)
    package["manual_review_flags"].extend(unit_reviews)
    package["source_coverage"]["legacy_adapter_truncations"] = adapter_truncations
    package["source_coverage"]["legacy_adapter_deduplications"] = adapter_deduplications
    package["source_coverage"]["legacy_unit_normalizations"] = unit_reviews
    package["source_coverage"]["text_quality_demotions"] = text_quality_demotions
    package["source_coverage"]["text_quality_summary"] = _text_quality_demotion_summary(text_quality_demotions)
    package["source_coverage"]["guidance_routing_reviews"] = guidance_routing_reviews
    return package


def _path_get(obj: Any, dotted_path: str) -> Any:
    current = obj
    for part in dotted_path.split("."):
        if isinstance(current, Mapping):
            if part not in current:
                return None
            current = current[part]
            continue
        if isinstance(current, list):
            try:
                current = current[int(part)]
            except (IndexError, ValueError):
                return None
            continue
        return None
    return current


def _indexed_collection_path(path: str) -> tuple[str, str] | None:
    parts = path.split(".")
    for idx, part in enumerate(parts):
        if part == "0" and idx > 0 and idx < len(parts) - 1:
            return ".".join(parts[:idx]), ".".join(parts[idx + 1 :])
    return None


def _field_is_populated(value: Any) -> bool:
    return isinstance(value, Mapping) and str(value.get("status") or "") == "populated" and value.get("value") not in (None, "")


def _field_source_ref(value: Any) -> str:
    if isinstance(value, Mapping):
        return str(value.get("source_ref") or "")
    return ""


def _count_values_for_binding(package: Mapping[str, Any], binding: Mapping[str, Any]) -> tuple[int, int, list[str]]:
    row_schema = binding.get("row_schema") if isinstance(binding.get("row_schema"), list) else []
    normalized_field = str(binding.get("normalized_field") or "")
    refs: set[str] = set()
    populated = 0
    available_rows = 0

    if str(binding.get("planning_mode") or "") == "pivot_rows":
        selector = binding.get("row_selector") if isinstance(binding.get("row_selector"), Mapping) else {}
        collection = _path_get(package, str(selector.get("source_path") or ""))
        value_field = str(binding.get("value_field") or "")
        if isinstance(collection, list) and value_field:
            for item in collection:
                if not isinstance(item, Mapping):
                    continue
                value = _path_get(item, value_field)
                if _field_is_populated(value):
                    populated += 1
                    available_rows += 1
                    ref = _field_source_ref(value)
                    if ref:
                        refs.add(ref)
        return populated, available_rows, sorted(refs)

    if row_schema:
        collection_path = _collection_path_for_row_schema(binding)
        collection = _path_get(package, collection_path) if collection_path else None
        if isinstance(collection, list):
            for item in collection:
                if not isinstance(item, Mapping):
                    continue
                row_has_value = False
                for column in row_schema:
                    source_field = str(column.get("source_field") or "")
                    value = _path_get(item, source_field)
                    if _field_is_populated(value) or (not isinstance(value, Mapping) and value not in (None, "")):
                        populated += 1
                        row_has_value = True
                    ref = _field_source_ref(value)
                    if ref:
                        refs.add(ref)
                if row_has_value:
                    available_rows += 1
        return populated, available_rows, sorted(refs)

    parsed = _indexed_collection_path(normalized_field)
    if parsed:
        collection_path, field_path = parsed
        collection = _path_get(package, collection_path)
        if isinstance(collection, list):
            for item in collection:
                value = _path_get(item, field_path)
                if _field_is_populated(value):
                    populated += 1
                    refs.add(_field_source_ref(value))
            return populated, populated, sorted(ref for ref in refs if ref)

    value = _path_get(package, normalized_field)
    if _field_is_populated(value):
        return 1, 1, [_field_source_ref(value)] if _field_source_ref(value) else []
    if isinstance(value, list):
        populated = len(value)
        return populated, populated, []
    return 0, 0, []


def _collection_path_for_row_schema(binding: Mapping[str, Any]) -> str:
    collection_path = str(binding.get("row_source") or "")
    if collection_path:
        return collection_path
    normalized_field = str(binding.get("normalized_field") or "")
    parsed = _indexed_collection_path(normalized_field)
    return parsed[0] if parsed else normalized_field


def _target_capacity(binding: Mapping[str, Any]) -> tuple[int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(str(binding.get("target") or "A1:A1"))
    return max_row - min_row + 1, max_col - min_col + 1


def build_binding_coverage_audit(
    package: Mapping[str, Any],
    binding_payload: Mapping[str, Any] | Sequence[Mapping[str, Any]],
    *,
    manifest: Mapping[str, Any],
    shell_path: Path | str = DEFAULT_SHELL,
    cached_plan: Mapping[str, Any] | BindingPlan | BindingPlanSnapshot | None = None,
) -> dict[str, Any]:
    """Calculate coverage from an independently reproduced authoritative plan.

    ``cached_plan`` is comparison-only.  Its type or self-contained digest never
    authorizes coverage and any difference from reproduction fails closed.
    """

    binding_map = (
        list(binding_payload.get("bindings") or [])
        if isinstance(binding_payload, Mapping)
        else list(binding_payload)
    )
    reproduced_plan = reproduce_binding_plan(
        package,
        binding_payload=binding_payload,
        manifest=manifest,
        shell_path=shell_path,
        expected_plan=cached_plan,
    )
    plan_payload = reproduced_plan.to_dict()
    planner_reports = {
        str(row.get("binding_id") or ""): row
        for row in plan_payload.get("bindings") or []
        if isinstance(row, Mapping)
    }
    planner_write_counts: dict[str, int] = defaultdict(int)
    for write in plan_payload.get("planned_writes") or []:
        if isinstance(write, Mapping):
            planner_write_counts[str(write.get("binding_id") or "")] += 1

    rows: list[dict[str, Any]] = []
    for binding in binding_map:
        values_available, rows_available, source_refs = _count_values_for_binding(package, binding)
        expected_rows, expected_cols = _target_capacity(binding)
        row_schema = binding.get("row_schema") if isinstance(binding.get("row_schema"), list) else []
        eligibility = inspect_binding_eligibility(package, binding)
        selected_rows = list(eligibility.get("selected_rows") or [])
        exclusions = list(eligibility.get("structured_exclusions") or [])
        planning_state = str(binding.get("planning_state") or "active")
        is_validation_output = str(binding.get("source_policy") or "") == "validation-output"
        if isinstance(binding.get("row_selector"), Mapping):
            planner_eligible_rows = len(selected_rows)
            planner_eligible_values = _count_selected_binding_values(selected_rows, binding)
        else:
            planner_eligible_rows = rows_available
            planner_eligible_values = values_available
        binding_id = str(binding.get("binding_id") or "")
        planner_report = planner_reports.get(binding_id, {})
        exact_planned_write_count = planner_write_counts.get(binding_id, 0)
        would_write = (
            planning_state == "active"
            and not is_validation_output
            and exact_planned_write_count > 0
        )
        reason = ""
        if not would_write:
            if planning_state != "active":
                reason = f"binding planning_state is {planning_state}"
            elif is_validation_output:
                reason = "validation output binding"
            elif values_available > 0 and planner_eligible_values == 0:
                reason = "raw normalized data exists but selector/pick/window produced no writable value"
            else:
                reason = "normalized field is absent or not populated"
        rows.append(
            {
                "binding_id": binding.get("binding_id", ""),
                "sheet": binding.get("sheet", ""),
                "section": binding.get("section", ""),
                "target": binding.get("target", ""),
                "normalized_field": binding.get("normalized_field", ""),
                "value_shape": binding.get("value_shape", ""),
                "required": bool(binding.get("required")),
                "row_schema_columns": [str(column.get("column_id") or "") for column in row_schema],
                "has_populated_data": values_available > 0,
                "number_of_values_available": values_available,
                "number_of_rows_available": rows_available,
                "number_of_rows_planner_eligible": planner_eligible_rows,
                "number_of_values_planner_eligible": planner_eligible_values,
                "planner_planned_write_count": exact_planned_write_count,
                "planner_capacity_used": int(planner_report.get("capacity_used") or 0),
                "planner_overflow_count": len(planner_report.get("overflow_rows") or []),
                "planner_structured_exclusion_count": len(planner_report.get("skipped_rows") or []),
                "number_of_rows_expected": expected_rows,
                "number_of_cells_expected": expected_rows * expected_cols,
                "source_ref_coverage": source_refs,
                "would_write_useful_output": would_write,
                "blank_reason": reason,
                "planning_state": planning_state,
                "structured_exclusion_count": len(exclusions),
                "structured_exclusion_reasons": sorted({str(row.get("reason") or "") for row in exclusions}),
                "selection_issue_count": len(eligibility.get("issues") or []),
            }
        )
    binding_document = binding_payload if isinstance(binding_payload, Mapping) else {"bindings": binding_map}
    return {
        "version": "0.3.0",
        "generator_version": "anf_legacy_adapter_audits/0.3.0",
        "source_package_content_sha256": _payload_sha256(package),
        "binding_contract_content_sha256": _payload_sha256(binding_document),
        "generated_at_utc": _now(),
        "ticker": "ANF",
        "planner_status": str(plan_payload.get("status") or "not_supplied"),
        "planner_total_write_count": int(plan_payload.get("planned_write_count") or len(plan_payload.get("planned_writes") or [])),
        "bindings": rows,
        "summary": {
            "binding_count": len(rows),
            "bindings_with_populated_data": sum(1 for row in rows if row["has_populated_data"]),
            "bindings_that_would_write_useful_output": sum(1 for row in rows if row["would_write_useful_output"]),
        },
    }


def _count_selected_binding_values(
    rows: Sequence[Mapping[str, Any]],
    binding: Mapping[str, Any],
) -> int:
    if str(binding.get("planning_mode") or "") == "pivot_rows":
        fields = [str(binding.get("value_field") or "")]
    else:
        fields = [
        str(column.get("source_field") or "")
        for column in binding.get("target_columns") or []
        if isinstance(column, Mapping) and column.get("source_field")
        ]
    if not fields:
        fields = [str(binding.get("source_field") or "")]
    count = 0
    for row in rows:
        for field in fields:
            value = _path_get(row, field)
            if _field_is_populated(value) or (not isinstance(value, Mapping) and value not in (None, "")):
                count += 1
    return count


def build_source_audit(package: Mapping[str, Any], *, data_root: Path, workbook_path: Path) -> dict[str, Any]:
    source_coverage = package.get("source_coverage", {})
    family_counts = source_coverage.get("family_counts", {}) if isinstance(source_coverage, Mapping) else {}
    section_sources = {
        "ticker_metadata": ["SEC company_tickers", "sec_cache/ANF/0001018840", "ANF_model.xlsx!SUMMARY"],
        "company_profile": ["ANF_model.xlsx!SUMMARY", "company profile configuration", "earnings release About section"],
        "quarterly_financials": ["ANF_model.xlsx!History_Q", "SEC/XBRL cache", "earnings release financial schedules"],
        "calculation_history": ["ANF_model.xlsx!History_Q projected as a period-keyed formula input ledger"],
        "annual_financials": ["ANF_model.xlsx!History_Q aggregated by fiscal_year", "annual reports", "earnings release annual schedules"],
        "debt_liquidity": ["ANF_model.xlsx!Leverage_Liquidity", "ANF_model.xlsx!History_Q", "ANF_model.xlsx!Slides_Debt_Profile"],
        "capital_returns": ["ANF_model.xlsx!History_Q", "earnings release capital allocation text"],
        "normalized_guidance": ["ANF_model.xlsx!Guidance_Normalized", "ANF_model.xlsx!Promise_Progress", "earnings releases", "transcripts"],
        "promise_progress": ["ANF_model.xlsx!Guidance_Normalized", "ANF_model.xlsx!Promise_Progress", "annual reports", "earnings releases", "transcripts"],
        "segments": ["ANF_model.xlsx!Slides_Segments", "earnings release segment tables", "presentation tables"],
        "operating_drivers": ["ANF_model.xlsx!operating_drivers_raw", "transcripts", "earnings presentations"],
        "quarter_notes": ["ANF_model.xlsx!Quarter_Notes", "ANF_model.xlsx!Quarter_Notes_Evidence"],
        "investment_case": ["ANF_model.xlsx!SUMMARY", "ANF_model.xlsx!ANF_Investment_Case_Data"],
        "valuation_outputs": ["explicit normalized valuation output builder (not available in the ANF legacy adapter fixture)"],
        "source_coverage": ["StockModelData/tickers/ANF", "StockModelData/sec_cache/ANF"],
        "mapping_gaps": ["docs/workbook_binding_map.json", "normalized package"],
        "manual_review_flags": ["pre-render validation", "mapping gap report"],
    }
    sections: list[dict[str, Any]] = []
    for section in REQUIRED_SECTIONS:
        populated_count = _section_populated_count(package.get(section))
        status_counts = _section_status_counts(package.get(section))
        review_count = _section_manual_review_count(package, section)
        sections.append(
            {
                "section": section,
                "available_source_candidates": section_sources[section],
                "source_backed_available": populated_count > 0 and section not in {"mapping_gaps", "manual_review_flags"},
                "profile_backed_available": section in {"ticker_metadata", "company_profile", "investment_case"},
                "legacy_workbook_derived_available": any("ANF_model.xlsx" in source for source in section_sources[section]),
                "missing_source": bool(status_counts.get("missing_source")) or (populated_count == 0 and section not in {"mapping_gaps", "manual_review_flags"}),
                "missing_mapping": bool(status_counts.get("missing_mapping")) or (section == "mapping_gaps" and bool(package.get("mapping_gaps"))),
                "parser_conflict": bool(status_counts.get("parser_conflict")),
                "manual_review_required": bool(review_count) or bool(status_counts.get("manual_review_required")),
                "populated_field_count": populated_count,
                "field_status_counts": status_counts,
                "manual_review_record_count": review_count,
            }
        )
    return {
        "version": "0.2.0",
        "generator_version": "anf_legacy_adapter_audits/0.2.0",
        "source_package_content_sha256": _payload_sha256(package),
        "generated_at_utc": _now(),
        "ticker": "ANF",
        "data_root": str(data_root),
        "legacy_workbook": str(workbook_path),
        "source_family_counts": family_counts,
        "sections": sections,
    }


def build_anf_text_quality_audit(package: Mapping[str, Any]) -> dict[str, Any]:
    audit = build_normalized_text_quality_audit(package)
    source_coverage = package.get("source_coverage", {}) if isinstance(package, Mapping) else {}
    demotions = (
        source_coverage.get("text_quality_demotions", [])
        if isinstance(source_coverage, Mapping) and isinstance(source_coverage.get("text_quality_demotions"), list)
        else []
    )
    demotion_summary = _text_quality_demotion_summary(demotions)
    visible_clean_by_section: dict[str, int] = defaultdict(int)
    visible_non_clean_by_section: dict[str, int] = defaultdict(int)
    for row in audit["rows"]:
        if not row.get("visible_ui"):
            continue
        section = str(row.get("section") or "unknown")
        if row.get("classification") == "clean_visible_ui":
            visible_clean_by_section[section] += 1
        else:
            visible_non_clean_by_section[section] += 1
    before_after: dict[str, dict[str, int]] = {}
    for section in sorted(set(visible_clean_by_section) | set(visible_non_clean_by_section) | set(demotion_summary["by_section"])):
        demoted = int(demotion_summary["by_section"].get(section, 0))
        after_clean = int(visible_clean_by_section.get(section, 0))
        before_after[section] = {
            "candidate_rows_before_filter": after_clean + demoted + int(visible_non_clean_by_section.get(section, 0)),
            "visible_clean_after_filter": after_clean,
            "visible_non_clean_after_filter": int(visible_non_clean_by_section.get(section, 0)),
            "demoted_before_render": demoted,
        }
    return {
        **audit,
        "ticker": "ANF",
        "generator_version": "anf_legacy_adapter_audits/0.2.0",
        "source_package_content_sha256": _payload_sha256(package),
        "generated_at_utc": _now(),
        "demotion_summary": demotion_summary,
        "before_after_summary": before_after,
        "demotions": list(demotions),
    }


def _section_populated_count(obj: Any) -> int:
    count = 0
    if isinstance(obj, Mapping):
        if _field_is_populated(obj):
            return 1
        for value in obj.values():
            count += _section_populated_count(value)
    elif isinstance(obj, list):
        for value in obj:
            count += _section_populated_count(value)
    return count


def _section_status_counts(obj: Any) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)

    def walk(value: Any) -> None:
        if isinstance(value, Mapping):
            status = str(value.get("status") or "")
            if status:
                counts[status] += 1
                return
            for child in value.values():
                walk(child)
        elif isinstance(value, list):
            for child in value:
                walk(child)

    walk(obj)
    return dict(sorted(counts.items()))


def _section_manual_review_count(package: Mapping[str, Any], section: str) -> int:
    rows = package.get("manual_review_flags") if isinstance(package.get("manual_review_flags"), list) else []
    if section == "manual_review_flags":
        return len(rows)
    if section == "mapping_gaps":
        gaps = package.get("mapping_gaps") if isinstance(package.get("mapping_gaps"), list) else []
        return len(gaps)
    count = 0
    for row in rows:
        if not isinstance(row, Mapping):
            continue
        row_section = str(row.get("section") or "")
        field = str(row.get("normalized_path") or row.get("field") or "")
        if row_section == section or field == section or field.startswith(section + "."):
            count += 1
    return count


def _payload_sha256(payload: Mapping[str, Any]) -> str:
    canonical = json.dumps(payload, sort_keys=True, ensure_ascii=False, separators=(",", ":"), default=str)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _markdown_source_audit(audit: Mapping[str, Any]) -> str:
    lines = [
        "# ANF Normalized Package Source Audit",
        "",
        "Read-only audit for the ANF shadow normalized-data package. This document does not define workbook rendering behavior.",
        "",
        f"- Generated: `{audit['generated_at_utc']}`",
        f"- Legacy workbook: `{audit['legacy_workbook']}`",
        "",
        "| Section | Classification | Source candidates | Populated fields |",
        "| --- | --- | --- | ---: |",
    ]
    for row in audit["sections"]:
        classes = [
            label
            for key, label in (
                ("source_backed_available", "source-backed available"),
                ("profile_backed_available", "profile-backed available"),
                ("legacy_workbook_derived_available", "legacy-workbook-derived available"),
                ("missing_source", "missing source"),
                ("missing_mapping", "missing mapping"),
                ("parser_conflict", "parser conflict"),
                ("manual_review_required", "manual review required"),
            )
            if row.get(key)
        ]
        lines.append(
            f"| `{row['section']}` | {', '.join(classes) or 'none'} | {'; '.join(row['available_source_candidates'])} | {row['populated_field_count']} |"
        )
    lines.extend(
        [
            "",
            "## Notes",
            "",
            "- ANF shadow data is read from saved source/workbook artifacts only.",
            "- Missing data remains a mapping gap or manual-review item; no generic filler text is introduced.",
            "- Real workbook rendering is intentionally out of scope for this pass.",
            "",
        ]
    )
    return "\n".join(lines)


def _markdown_binding_coverage(audit: Mapping[str, Any]) -> str:
    lines = [
        "# ANF Binding Coverage Audit",
        "",
        "Coverage check for how the ANF shadow normalized package maps to the current workbook binding map.",
        "",
        f"- Generated: `{audit['generated_at_utc']}`",
        f"- Bindings with populated data: `{audit['summary']['bindings_with_populated_data']}` / `{audit['summary']['binding_count']}`",
        f"- Bindings that would write useful output: `{audit['summary']['bindings_that_would_write_useful_output']}`",
        "",
        "| Binding | Sheet | Field | Values | Rows | Would write | Reason if blank |",
        "| --- | --- | --- | ---: | ---: | --- | --- |",
    ]
    for row in audit["bindings"]:
        lines.append(
            f"| `{row['binding_id']}` | `{row['sheet']}` | `{row['normalized_field']}` | {row['number_of_values_available']} | {row['number_of_rows_available']}/{row['number_of_rows_expected']} | {row['would_write_useful_output']} | {row['blank_reason']} |"
        )
    lines.extend(
        [
            "",
            "## Row Schema Observation",
            "",
            "Table-row bindings now expose row-schema columns in the JSON binding map. The ANF shadow package populates enough row-shaped data to audit whether future filler output would be useful, without creating an ANF workbook in this pass.",
            "",
        ]
    )
    return "\n".join(lines)


def _markdown_text_quality_audit(audit: Mapping[str, Any]) -> str:
    lines = [
        "# ANF Normalized Text Quality Audit",
        "",
        "Read-only text-quality audit for the ANF shadow normalized package. Non-renderable snippets are demoted to manual review/source coverage rather than copied into visible UI fields.",
        "",
        f"- Generated: `{audit['generated_at_utc']}`",
        f"- Audited text rows: `{audit['row_count']}`",
        f"- Non-clean visible rows after filtering: `{audit['non_clean_visible_count']}`",
        f"- Demoted rows before render: `{audit['demotion_summary']['total_demoted']}`",
        "",
        "## Before / After By Section",
        "",
        "| Section | Candidate rows before filter | Visible clean after filter | Visible non-clean after filter | Demoted before render |",
        "| --- | ---: | ---: | ---: | ---: |",
    ]
    for section, row in audit["before_after_summary"].items():
        lines.append(
            f"| `{section}` | {row['candidate_rows_before_filter']} | {row['visible_clean_after_filter']} | {row['visible_non_clean_after_filter']} | {row['demoted_before_render']} |"
        )
    lines.extend(
        [
            "",
            "## Demotions",
            "",
            "| Field | Classification | Source | Action | Excerpt |",
            "| --- | --- | --- | --- | --- |",
        ]
    )
    displayed_demotions = list(audit["demotions"][:120])
    for row in displayed_demotions:
        excerpt = str(row.get("original_excerpt") or "").replace("|", "\\|")
        lines.append(
            f"| `{row.get('field', '')}` | `{row.get('classification', '')}` | `{row.get('source_ref', '')}` | {row.get('suggested_action', '')} | {excerpt} |"
        )
    omitted_count = len(audit["demotions"]) - len(displayed_demotions)
    if omitted_count:
        lines.append(f"| ... | ... | ... | ... | {omitted_count} additional demotions omitted from markdown; see JSON. |")
    lines.extend(
        [
            "",
            "## Policy",
            "",
            "- Clean source-backed text may remain in visible normalized fields.",
            "- Boilerplate, legal text, governance/compensation snippets, definitions, source headers, fragments, and overlong snippets stay audit-only/manual-review until normalized explicitly.",
            "- This audit does not create or fill any workbook.",
            "",
        ]
    )
    return "\n".join(lines)


def build_anf_shadow_outputs(
    *,
    data_root: Path,
    workbook_path: Path,
    output_dir: Path,
    docs_dir: Path,
    binding_map_path: Path | None = None,
) -> dict[str, Path]:
    binding_path = binding_map_path or (REPO_ROOT / "docs" / "workbook_binding_map.json")
    binding_payload = _load_json(binding_path)
    binding_map = list(binding_payload.get("bindings") or [])
    package = build_anf_normalized_package(data_root=data_root, workbook_path=workbook_path)
    raw_mapping_report = build_mapping_gap_report(package, binding_map, ticker="ANF")
    mapping_gaps = [
        gap
        for gap in raw_mapping_report.get("gaps", [])
        if str(gap.get("source_policy") or "") != "validation-output"
    ]
    mapping_report = {
        **raw_mapping_report,
        "gap_count": len(mapping_gaps),
        "gaps": mapping_gaps,
        "excluded_validation_output_gap_count": len(raw_mapping_report.get("gaps", [])) - len(mapping_gaps),
    }
    validation_issues = validate_normalized_company_data(package, binding_map=binding_map, promotion_requested=False)
    validation_report = {
        "ticker": "ANF",
        "promotion_requested": False,
        "issue_count": len(validation_issues),
        "issues": [issue.to_dict() for issue in validation_issues],
        "text_quality_demotion_count": len(
            package.get("source_coverage", {}).get("text_quality_demotions", [])
        ),
    }

    package_out = dict(package)
    package_out["mapping_gaps"] = mapping_report["gaps"]
    package_out["manual_review_flags"] = list(package.get("manual_review_flags", [])) + validation_report["issues"]
    source_audit = build_source_audit(package_out, data_root=data_root, workbook_path=workbook_path)
    manifest = _load_json(DEFAULT_MANIFEST)
    binding_coverage = build_binding_coverage_audit(
        package_out,
        binding_payload,
        manifest=manifest,
        shell_path=DEFAULT_SHELL,
    )
    text_quality_audit = build_anf_text_quality_audit(package_out)

    paths = {
        "package": output_dir / "ANF_normalized_data_package.json",
        "mapping_gaps": output_dir / "ANF_mapping_gaps_report.json",
        "validation": output_dir / "ANF_content_validation_report.json",
        "source_audit_json": docs_dir / "anf_normalized_package_source_audit.json",
        "source_audit_md": docs_dir / "anf_normalized_package_source_audit.md",
        "binding_coverage_json": docs_dir / "anf_binding_coverage_audit.json",
        "binding_coverage_md": docs_dir / "anf_binding_coverage_audit.md",
        "text_quality_json": docs_dir / "anf_normalized_text_quality_audit.json",
        "text_quality_md": docs_dir / "anf_normalized_text_quality_audit.md",
    }
    _write_json(paths["package"], package_out)
    _write_json(paths["mapping_gaps"], mapping_report)
    _write_json(paths["validation"], validation_report)
    _write_json(paths["source_audit_json"], source_audit)
    _write_text(paths["source_audit_md"], _markdown_source_audit(source_audit))
    _write_json(paths["binding_coverage_json"], binding_coverage)
    _write_text(paths["binding_coverage_md"], _markdown_binding_coverage(binding_coverage))
    _write_json(paths["text_quality_json"], text_quality_audit)
    _write_text(paths["text_quality_md"], _markdown_text_quality_audit(text_quality_audit))
    return paths


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-root", type=Path, default=_default_data_root())
    parser.add_argument("--workbook", type=Path, default=None)
    parser.add_argument("--output-dir", type=Path, default=None)
    parser.add_argument("--docs-dir", type=Path, default=REPO_ROOT / "docs")
    parser.add_argument("--binding-map", type=Path, default=REPO_ROOT / "docs" / "workbook_binding_map.json")
    args = parser.parse_args(argv)

    data_root = args.data_root.expanduser().resolve()
    workbook_path = (args.workbook or _default_workbook_path(data_root)).expanduser().resolve()
    output_dir = (args.output_dir or _default_output_dir(data_root)).expanduser().resolve()
    docs_dir = args.docs_dir.expanduser().resolve()
    paths = build_anf_shadow_outputs(
        data_root=data_root,
        workbook_path=workbook_path,
        output_dir=output_dir,
        docs_dir=docs_dir,
        binding_map_path=args.binding_map,
    )
    for label, path in paths.items():
        print(f"{label}: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
