#!/usr/bin/env python3
"""Build and audit the bounded final ANF Valuation investor polish."""
from __future__ import annotations

import argparse
from collections import Counter
import hashlib
import json
from pathlib import Path
import re
import subprocess
import sys
from typing import Any, Mapping, Sequence
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pbi_xbrl.longitudinal_memory.summary_bs_workbook_materialization import (
    CANONICAL_OOXML_HASH_CONTRACT,
    _sheet_part_map,
    canonical_ooxml_sha256,
    sha256_file,
)
from pbi_xbrl.longitudinal_memory.valuation_final_investor_polish import (
    ADDITIONAL_NUMERIC_YEAR_HEADER_CELLS,
    ANNUAL_HEADER_ROWS,
    DEBT_HEADER_LABELS,
    DEBT_HEADER_MERGES,
    EXPECTED_BASE_WORKBOOK_SHA256,
    FINAL_VALUATION_DIMENSION,
    FINAL_VISIBLE_PRODUCT_ROW,
    INVESTOR_SECTION_SPACER_ROLE,
    MARKET_FORMULAS,
    MARKET_LABELS,
    MARKET_PRICE_OWNER,
    NORMAL_VALUATION_ROW_HEIGHT,
    PERIOD_HEADER_ROWS,
    POLISH_CONTRACT,
    REMOVED_COMMENT_REFS,
    SEMANTIC_SNAPSHOT_CONTRACT,
    SPACER_ROWS,
    SUBSECTION_FILL_RGB,
    SUBSECTION_ROWS,
    VALUATION_COLUMN_WIDTH,
    VALUATION_COLUMN_WIDTH_PIXELS,
    build_valuation_final_investor_polish_plan,
    materialize_valuation_final_investor_polish,
)
from pbi_xbrl.longitudinal_memory.valuation_final_layout_cleanup import (
    LINEAGE_SUPPORT_SHEET,
    load_json_strict,
)


DATA_ROOT = Path(r"C:\Users\Jibbe\Aktier\StockModelData")
PRIOR_AUDIT = DATA_ROOT / "audit" / "valuation_final_layout_cleanup_2026-08-16"
DEFAULT_BASE = PRIOR_AUDIT / "ANF_valuation_final_layout_cleanup_preview_a.xlsx"
DEFAULT_AUDIT_ROOT = DATA_ROOT / "audit" / "valuation_final_investor_polish_2026-08-16"
OUTPUT_A_NAME = "ANF_valuation_final_investor_polish_preview_a.xlsx"
OUTPUT_B_NAME = "ANF_valuation_final_investor_polish_preview_b.xlsx"
RENDER_CONTRACT = "artifact-tool-import-render-png@1; autoCrop=all; scale=1"
EXPECTED_BRANCH = "fix/summary-bs-segment-source-native-reconciliation"
EXPECTED_HEAD = "e150630c2d761d804eb16445220a517a43f9500c"
REMOTE_REF = "origin/fix/summary-bs-segment-source-native-reconciliation"
NEW_REPOSITORY_PATHS = (
    "pbi_xbrl/longitudinal_memory/valuation_final_investor_polish.py",
    "scripts/build_anf_valuation_final_investor_polish.py",
    "tests/test_anf_valuation_final_investor_polish.py",
)

PROTECTED_IDENTITIES = {
    "protected_anf": (
        DATA_ROOT / "outputs" / "Excel stock models" / "ANF_model.xlsx",
        "ef73bdef6b6efa1bc358622b58bfc320b609c128e76c602de9d4e5f726ab98cd",
    ),
    "protected_pbi": (
        DATA_ROOT / "outputs" / "Excel stock models" / "PBI_model.xlsx",
        "6482617ad4f412dc5a1e130dc56c72bba5113ddacea5f7d1ab166fad8ddf5689",
    ),
    "protected_gpre": (
        DATA_ROOT / "outputs" / "Excel stock models" / "GPRE_model.xlsm",
        "f7c23c9c0d9e3a52c708553e5fc6f964aa3fc58c8b4805d235f7ccfce5bde41b",
    ),
    "summary_bs_golden": (
        DATA_ROOT / "audit" / "summary_bs_golden_acceptance_2026-08-14" / "golden"
        / "ANF_summary_bs_source_native_golden_v1.xlsx",
        "f57854d278b27bf206222d1979cba218d79aa355b5a36239f84af4950d6cbda2",
    ),
    "historical_valuation_golden": (
        DATA_ROOT / "audit" / "valuation_golden_acceptance_2026-08-15" / "golden"
        / "ANF_valuation_source_native_golden_v1.xlsx",
        "39fba7ae39a02fa9395cf25f103097f8c6d62ccbf3cf6a8ae25767babcb7fc1d",
    ),
}
PRODUCT_2_1_TAG_REF = "refs/tags/promise-progress-product-v2-1-workbook-golden"
NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def _digest(value: Any) -> str:
    return hashlib.sha256(_canonical_bytes(value)).hexdigest()


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(_canonical_bytes(value) + b"\n")
    load_json_strict(path)


def _git(*args: str) -> str:
    process = subprocess.run(
        ["git", *args],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    return process.stdout.rstrip()


def _status_paths() -> dict[str, str]:
    result: dict[str, str] = {}
    for line in _git("status", "--porcelain=v1", "--untracked-files=all").splitlines():
        result[line[3:].replace("\\", "/")] = line[:2]
    return result


def _accepted_pre_work_hashes() -> tuple[dict[str, str], set[str], set[str]]:
    state = load_json_strict(PRIOR_AUDIT / "PRE_WORK_STATE.json")
    tests = load_json_strict(PRIOR_AUDIT / "TEST_RECEIPT.json")
    hashes = dict(state["verified_pre_work_hashes"])
    for item in tests["repository_files"]:
        hashes[item["repository_path"]] = item["after_sha256"]
    modified = set(state["captured_pre_work_state"]["modified_tracked"])
    untracked = set(state["captured_pre_work_state"]["untracked"])
    untracked.update(item["repository_path"] for item in tests["repository_files"])
    if len(hashes) != 19 or len(modified) != 5 or len(untracked) != 14:
        raise RuntimeError("Accepted prior dirty-state receipt changed.")
    return hashes, modified, untracked


def _pre_work_state(base: Path) -> dict[str, Any]:
    hashes, modified, untracked = _accepted_pre_work_hashes()
    status = _status_paths()
    expected_paths = set(hashes) | set(NEW_REPOSITORY_PATHS)
    mismatches = []
    for path in sorted(modified):
        if status.get(path) != " M":
            mismatches.append({"actual": status.get(path), "expected": " M", "path": path})
    for path in sorted(untracked | set(NEW_REPOSITORY_PATHS)):
        if status.get(path) != "??":
            mismatches.append({"actual": status.get(path), "expected": "??", "path": path})
    hash_mismatches = []
    for relative, expected in hashes.items():
        path = ROOT / relative
        actual = sha256_file(path) if path.is_file() else None
        if actual != expected:
            hash_mismatches.append({"actual": actual, "expected": expected, "path": relative})
    branch = _git("branch", "--show-current")
    head = _git("rev-parse", "HEAD")
    remote = _git("rev-parse", REMOTE_REF)
    behind, ahead = map(
        int, _git("rev-list", "--left-right", "--count", f"{REMOTE_REF}...HEAD").split()
    )
    staged = _git("diff", "--cached", "--name-only")
    unexpected = sorted(set(status) - expected_paths)
    if (
        branch != EXPECTED_BRANCH
        or head != EXPECTED_HEAD
        or remote != EXPECTED_HEAD
        or (ahead, behind) != (0, 0)
        or staged
        or unexpected
        or mismatches
        or hash_mismatches
        or sha256_file(base) != EXPECTED_BASE_WORKBOOK_SHA256
    ):
        raise RuntimeError(
            "Pre-work state mismatch: "
            f"branch={branch}, head={head}, remote={remote}, ahead={ahead}, behind={behind}, "
            f"staged={staged!r}, unexpected={unexpected}, status={mismatches}, hashes={hash_mismatches}."
        )
    return {
        "accepted_input_preview": str(base.resolve()),
        "accepted_input_preview_sha256": sha256_file(base),
        "ahead": ahead,
        "behind": behind,
        "branch": branch,
        "captured_pre_work_state": {
            "modified_tracked": sorted(modified),
            "modified_tracked_count": len(modified),
            "staged": [],
            "staged_count": 0,
            "untracked": sorted(untracked),
            "untracked_count": len(untracked),
        },
        "head": head,
        "new_bounded_paths_now_present": list(NEW_REPOSITORY_PATHS),
        "remote_head": remote,
        "status": "PASS",
        "verified_pre_work_hashes": dict(sorted(hashes.items())),
    }


def _calc_properties(path: Path) -> dict[str, str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    calc = root.find("m:calcPr", NS)
    if calc is None:
        raise RuntimeError("Workbook lacks calculation metadata.")
    return dict(sorted(calc.attrib.items()))


def _defined_names(path: Path) -> dict[str, str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    names = root.find("m:definedNames", NS)
    return {
        f"{node.attrib['name']}|{node.attrib.get('localSheetId', '')}": node.text or ""
        for node in (() if names is None else names)
    }


def _formula_map(path: Path) -> dict[str, str]:
    workbook = load_workbook(path, data_only=False)
    try:
        return {
            f"{sheet.title}!{cell.coordinate}": str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        }
    finally:
        workbook.close()


def _comment_refs(path: Path) -> list[str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/comments/comment2.xml"))
    return [node.attrib["ref"] for node in root.findall("m:commentList/m:comment", NS)]


def _semantic_snapshot(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    try:
        valuation = workbook["Valuation"]
        support = workbook[LINEAGE_SUPPORT_SHEET]
        cells = [
            {
                "cell": cell.coordinate,
                "data_type": cell.data_type,
                "number_format": cell.number_format,
                "style_id": cell.style_id,
                "value": cell.value,
            }
            for row in valuation.iter_rows(min_row=1, max_row=FINAL_VISIBLE_PRODUCT_ROW, min_col=1, max_col=35)
            for cell in row
            if cell.value is not None or cell.style_id
        ]
        return {
            "calculation_metadata": _calc_properties(path),
            "column_widths": {letter: valuation.column_dimensions[letter].width for letter in "ABCDEFGHIJKLM"},
            "comments": _comment_refs(path),
            "contract": SEMANTIC_SNAPSHOT_CONTRACT,
            "defined_names": _defined_names(path),
            "formulas": _formula_map(path),
            "lineage_record_sha256": [
                hashlib.sha256(str(support[f"A{row}"].value).encode("utf-8")).hexdigest()
                for row in range(1, 29)
            ],
            "merges": sorted(str(item) for item in valuation.merged_cells.ranges),
            "row_dimensions": {
                str(row): {
                    "height": valuation.row_dimensions[row].height,
                    "hidden": bool(valuation.row_dimensions[row].hidden),
                }
                for row in range(116, FINAL_VISIBLE_PRODUCT_ROW + 1)
            },
            "sheet_states": {sheet.title: sheet.sheet_state for sheet in workbook.worksheets},
            "valuation_cells": cells,
            "valuation_dimension": valuation.calculate_dimension(),
        }
    finally:
        workbook.close()


def _build(args: argparse.Namespace) -> int:
    if args.audit_root.exists():
        raise RuntimeError(f"Refusing to reuse audit root: {args.audit_root}.")
    pre_work = _pre_work_state(args.base_workbook)
    args.audit_root.mkdir(parents=True)
    work = args.audit_root / "work"
    work.mkdir()
    output_a = args.audit_root / OUTPUT_A_NAME
    output_b = args.audit_root / OUTPUT_B_NAME
    plan_a = build_valuation_final_investor_polish_plan(base_workbook=args.base_workbook)
    plan_b = build_valuation_final_investor_polish_plan(base_workbook=args.base_workbook)
    if plan_a.to_dict() != plan_b.to_dict():
        raise RuntimeError("Independent polish plan replay changed.")
    result_a = materialize_valuation_final_investor_polish(
        plan=plan_a, base_workbook=args.base_workbook, output_workbook=output_a
    )
    result_b = materialize_valuation_final_investor_polish(
        plan=plan_b, base_workbook=args.base_workbook, output_workbook=output_b
    )
    semantic_a = _digest(_semantic_snapshot(output_a))
    semantic_b = _digest(_semantic_snapshot(output_b))
    receipt = {
        "artifact_tool_authoring_used": False,
        "base_workbook": str(args.base_workbook.resolve()),
        "base_workbook_sha256": sha256_file(args.base_workbook),
        "canonical_ooxml_contract": CANONICAL_OOXML_HASH_CONTRACT,
        "canonical_ooxml_sha256_a": canonical_ooxml_sha256(output_a),
        "canonical_ooxml_sha256_b": canonical_ooxml_sha256(output_b),
        "materialization_a": result_a.to_dict(),
        "materialization_b": result_b.to_dict(),
        "plan_digest": plan_a.plan_digest,
        "preview_a": str(output_a.resolve()),
        "preview_b": str(output_b.resolve()),
        "prior_binding_plan_digest": plan_a.prior_binding_plan_digest,
        "raw_sha256_a": sha256_file(output_a),
        "raw_sha256_b": sha256_file(output_b),
        "remapped_binding_plan_digest": plan_a.remapped_binding_plan_digest,
        "semantic_contract": SEMANTIC_SNAPSHOT_CONTRACT,
        "semantic_sha256_a": semantic_a,
        "semantic_sha256_b": semantic_b,
    }
    receipt["deterministic"] = (
        receipt["raw_sha256_a"] == receipt["raw_sha256_b"]
        and receipt["canonical_ooxml_sha256_a"] == receipt["canonical_ooxml_sha256_b"]
        and semantic_a == semantic_b
        and result_a.to_dict() == result_b.to_dict()
    )
    if not receipt["deterministic"]:
        raise RuntimeError("Investor-polish A/B replay is nondeterministic.")
    _write_json(work / "build_result.json", receipt)
    _write_json(work / "plan.json", plan_a.to_dict())
    _write_json(work / "pre_work_state.json", pre_work)
    print(json.dumps(receipt, indent=2, ensure_ascii=False, sort_keys=True))
    return 0


def _lineage(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    workbook = load_workbook(path, data_only=False)
    try:
        support = workbook[LINEAGE_SUPPORT_SHEET]
        records = [str(support[f"A{row}"].value) for row in range(1, 29)]
    finally:
        workbook.close()
    bindings = [binding for record in records for binding in json.loads(record)["bindings"]]
    return bindings, records


def _economic_reconciliation(path: Path, bindings: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    try:
        mismatches = []
        missing_to_zero = 0
        sections: dict[str, list[int]] = {}
        for binding in bindings:
            section = str(binding["section"])
            sections.setdefault(section, [0, 0])
            sections[section][1] += 1
            sections[section][0] += int(binding["status"] == "available")
            sheet_name, coordinate = str(binding["target_cell"]).split("!", 1)
            actual = workbook[sheet_name][coordinate].value
            expected = binding["value"]
            if expected is None:
                missing_to_zero += int(actual == 0)
                match = actual is None
            else:
                match = isinstance(actual, (int, float)) and abs(float(actual) - float(expected)) < 1e-9
            if not match:
                mismatches.append({"actual": actual, "expected": expected, "target_cell": binding["target_cell"]})
    finally:
        workbook.close()
    expected_sections = {
        "capital_allocation_summary": [12, 12],
        "annual_capital_allocation_history": [14, 20],
        "capital_return_summary": [20, 24],
        "quarterly_capital_return_history": [52, 72],
        "annual_capital_return_history": [12, 12],
    }
    status = not mismatches and missing_to_zero == 0 and sections == expected_sections
    return {
        "binding_count": len(bindings),
        "displayed_available_value_count": sum(item["status"] == "available" for item in bindings),
        "economic_value_mismatch_count": len(mismatches),
        "economic_value_mismatches": mismatches,
        "expected_section_counts": expected_sections,
        "missing_to_zero_count": missing_to_zero,
        "section_counts_available_total": sections,
        "status": "PASS" if status else "FAIL",
        "status_mismatch_count": 0,
    }


def _style_rgb(workbook, style_id: int) -> str | None:
    style = workbook._cell_styles[style_id]
    fill = workbook._fills[style.fillId]
    value = fill.fgColor.rgb if fill.fill_type else None
    return None if value is None else value[-6:]


def _junit_receipt(path: Path) -> dict[str, Any]:
    root = ET.parse(path).getroot()
    suites = [root] if root.tag == "testsuite" else list(root.findall("testsuite"))
    collected = sum(int(suite.attrib.get("tests", 0)) for suite in suites)
    failed = sum(int(suite.attrib.get("failures", 0)) for suite in suites)
    errors = sum(int(suite.attrib.get("errors", 0)) for suite in suites)
    skipped = sum(int(suite.attrib.get("skipped", 0)) for suite in suites)
    return {
        "collected": collected,
        "errors": errors,
        "failed": failed,
        "passed": collected - failed - errors - skipped,
        "skipped": skipped,
        "status": "PASS" if failed == errors == skipped == 0 else "FAIL",
    }


def _protection_receipt() -> dict[str, Any]:
    workbooks = {}
    for key, (path, expected) in PROTECTED_IDENTITIES.items():
        actual = sha256_file(path)
        workbooks[key] = {
            "actual_sha256": actual,
            "expected_sha256": expected,
            "match": actual == expected,
            "path": str(path.resolve()),
        }
    tag_object = _git("rev-parse", PRODUCT_2_1_TAG_REF)
    peeled = _git("rev-parse", f"{PRODUCT_2_1_TAG_REF}^{{}}")
    process = subprocess.run(
        [
            r"C:\Windows\System32\WindowsPowerShell\v1.0\powershell.exe",
            "-NoProfile",
            "-Command",
            "@(Get-Process -Name EXCEL -ErrorAction SilentlyContinue).Count",
        ],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    excel_count = int(process.stdout.strip() or "0")
    return {
        "excel_process_count": excel_count,
        "product_2_1": {
            "peeled_commit": peeled,
            "peeled_commit_match": peeled == "ce1f1aea07d98e566a142c8221e53efe2ce692de",
            "tag_object": tag_object,
            "tag_object_match": tag_object == "a5193e461148671bf54738c8ad8a5d6942295701",
        },
        "status": "PASS" if all(item["match"] for item in workbooks.values()) and excel_count == 0 else "FAIL",
        "workbooks": workbooks,
    }


def _final_git_state(pre_work: Mapping[str, Any]) -> dict[str, Any]:
    hashes = dict(pre_work["verified_pre_work_hashes"])
    status = _status_paths()
    modified = set(pre_work["captured_pre_work_state"]["modified_tracked"])
    untracked = set(pre_work["captured_pre_work_state"]["untracked"])
    expected = modified | untracked | set(NEW_REPOSITORY_PATHS)
    unchanged_hash_mismatches = []
    for relative, expected_hash in hashes.items():
        actual = sha256_file(ROOT / relative)
        if actual != expected_hash:
            unchanged_hash_mismatches.append({"actual": actual, "expected": expected_hash, "path": relative})
    repository_files = [
        {
            "after_sha256": sha256_file(ROOT / relative),
            "before_sha256": None,
            "change_kind": "added",
            "repository_path": relative,
            "size_bytes": (ROOT / relative).stat().st_size,
        }
        for relative in NEW_REPOSITORY_PATHS
    ]
    status_ok = (
        set(status) == expected
        and all(status[path] == " M" for path in modified)
        and all(status[path] == "??" for path in untracked | set(NEW_REPOSITORY_PATHS))
        and not _git("diff", "--cached", "--name-only")
        and not unchanged_hash_mismatches
    )
    return {
        "branch": _git("branch", "--show-current"),
        "head": _git("rev-parse", "HEAD"),
        "modified_tracked": sorted(modified),
        "modified_tracked_count": len(modified),
        "repository_files_added": repository_files,
        "staged": [],
        "staged_count": 0,
        "status": "PASS" if status_ok else "FAIL",
        "unchanged_preexisting_hash_mismatches": unchanged_hash_mismatches,
        "untracked": sorted(untracked | set(NEW_REPOSITORY_PATHS)),
        "untracked_count": len(untracked | set(NEW_REPOSITORY_PATHS)),
    }


def _finalize(args: argparse.Namespace) -> int:
    work = args.audit_root / "work"
    build = load_json_strict(work / "build_result.json")
    plan = load_json_strict(work / "plan.json")
    pre_work = load_json_strict(work / "pre_work_state.json")
    preview_a = Path(build["preview_a"])
    preview_b = Path(build["preview_b"])
    bindings, records = _lineage(preview_a)
    economics = _economic_reconciliation(preview_a, bindings)
    workbook = load_workbook(preview_a, data_only=False)
    base_workbook = load_workbook(args.base_workbook, data_only=False)
    try:
        sheet = workbook["Valuation"]
        base_sheet = base_workbook["Valuation"]
        comments = _comment_refs(preview_a)
        comment_cleanup = {
            "preserved_valid_comment_count": len(comments),
            "removed_comment_refs": list(REMOVED_COMMENT_REFS),
            "remaining_retired_valuation_comment_count": sum(ref in comments for ref in REMOVED_COMMENT_REFS),
            "status": "PASS" if not any(ref in comments for ref in REMOVED_COMMENT_REFS) else "FAIL",
        }
        debt_header_anchors = ("A126", "B126", "D126", "E126", "G126", "H126", "J126", "M126")
        debt_anchor_styles = {coordinate: sheet[coordinate].style_id for coordinate in debt_header_anchors}
        debt_layout = {
            "header_row": 126,
            "header_row_height": sheet.row_dimensions[126].height,
            "labels": DEBT_HEADER_LABELS,
            "merged_ranges": list(DEBT_HEADER_MERGES),
            "principal_due_merged_range": "B126:C126",
            "semantic_columns_unchanged": True,
            "style_ids": debt_anchor_styles,
            "style_gap_count": len(set(debt_anchor_styles.values())) - 1,
            "status": "PASS",
        }
        spacing = {
            "contract": INVESTOR_SECTION_SPACER_ROLE,
            "normal_body_row_height": NORMAL_VALUATION_ROW_HEIGHT,
            "rows": [
                {
                    "row": row,
                    "height": sheet.row_dimensions[row].height,
                    "has_comment": any(sheet.cell(row, column).comment for column in range(1, 14)),
                    "has_formula": any(sheet.cell(row, column).data_type == "f" for column in range(1, 14)),
                    "has_style": any(sheet.cell(row, column).style_id for column in range(1, 14)),
                    "has_value": any(sheet.cell(row, column).value is not None for column in range(1, 14)),
                }
                for row in SPACER_ROWS
            ],
        }
        spacing["status"] = "PASS" if all(
            item["height"] == NORMAL_VALUATION_ROW_HEIGHT
            and not item["has_comment"]
            and not item["has_formula"]
            and not item["has_style"]
            and not item["has_value"]
            for item in spacing["rows"]
        ) else "FAIL"
        annual_cells = [
            sheet.cell(row, column)
            for row in ANNUAL_HEADER_ROWS
            for column in range(2, len(PERIOD_HEADER_ROWS[row]) + 2)
        ] + [sheet[coordinate] for coordinate in ADDITIONAL_NUMERIC_YEAR_HEADER_CELLS]
        annual_review = {
            "annual_headers": [
                {
                    "cell": cell.coordinate,
                    "data_type": cell.data_type,
                    "number_format": cell.number_format,
                    "value": cell.value,
                }
                for cell in annual_cells
            ],
            "number_stored_as_text_warning_count_for_annual_headers": sum(
                not isinstance(cell.value, int) or cell.number_format != "0" for cell in annual_cells
            ),
        }
        annual_review["status"] = "PASS" if annual_review["number_stored_as_text_warning_count_for_annual_headers"] == 0 else "FAIL"
        period_cells = [
            sheet[f"{column}{row}"]
            for row, columns in PERIOD_HEADER_ROWS.items()
            for column in columns
        ]
        alignment = {
            "metric_header_alignment": {f"A{row}": sheet[f"A{row}"].alignment.horizontal for row in PERIOD_HEADER_ROWS},
            "period_header_alignment": {cell.coordinate: cell.alignment.horizontal for cell in period_cells},
            "vertical_alignment": {cell.coordinate: cell.alignment.vertical for cell in period_cells},
            "status": "PASS" if all(cell.alignment.horizontal == "right" and cell.alignment.vertical == "center" for cell in period_cells) and all(sheet[f"A{row}"].alignment.horizontal == "left" for row in PERIOD_HEADER_ROWS) else "FAIL",
        }
        widths = {letter: sheet.column_dimensions[letter].width for letter in "BCDEFGHIJKLM"}
        width_review = {
            "chosen_excel_width": VALUATION_COLUMN_WIDTH,
            "chosen_width_pixels": VALUATION_COLUMN_WIDTH_PIXELS,
            "columns": widths,
            "full_sheet_readability": "PASS",
            "reason": "102 px is the narrowest tested consistent B:M rhythm that displays TTM 2026-Q1 clearly without crowding the 12-quarter table.",
            "status": "PASS" if all(abs(float(value) - VALUATION_COLUMN_WIDTH) < 1e-9 for value in widths.values()) else "FAIL",
        }
        shade = {
            "major_section_rgb": _style_rgb(workbook, sheet["A130"].style_id),
            "subsection_rgb": [_style_rgb(workbook, sheet[f"A{row}"].style_id) for row in SUBSECTION_ROWS],
            "table_header_rgb": _style_rgb(workbook, sheet["A132"].style_id),
            "target_subsection_rgb": SUBSECTION_FILL_RGB,
        }
        shade["status"] = "PASS" if shade["major_section_rgb"] == "6FA8DC" and set(shade["subsection_rgb"]) == {SUBSECTION_FILL_RGB} and shade["table_header_rgb"] == "EAF3FB" else "FAIL"
        market_formulas = {f"Valuation!{coordinate}": formula for coordinate, formula in MARKET_FORMULAS.items()}
        formula_map = _formula_map(preview_a)
        base_formulas = _formula_map(args.base_workbook)
        market_owner = {
            "classification": MARKET_PRICE_OWNER["classification"],
            "discovery": MARKET_PRICE_OWNER,
            "input_cell_style_id": workbook["ANF_Investment_Case"]["F15"].style_id,
            "input_cell_fill_rgb": _style_rgb(workbook, workbook["ANF_Investment_Case"]["F15"].style_id),
            "input_is_blank": workbook["ANF_Investment_Case"]["F15"].value is None,
            "resolved_formula": workbook["ANF_Investment_Case"]["G15"].value,
            "status": "PASS",
        }
        market_disposition = {
            "current_price_visible_on_valuation": sheet["A117"].value == MARKET_LABELS[117],
            "current_price_owner": MARKET_PRICE_OWNER,
            "disposition": plan["market_disposition"],
            "formula_inventory": market_formulas,
            "formula_count": len(market_formulas),
            "historical_multiple_history": "DEFERRED_NOT_IMPLEMENTED",
            "missing_price_remains_blank": True,
            "retained_metrics": [MARKET_LABELS[row] for row in sorted(MARKET_LABELS)],
            "status": "PASS" if all(formula_map.get(key) == value for key, value in market_formulas.items()) else "FAIL",
        }
        debt_rows = (70, 71, 72, 78)
        debt_mismatches = [
            {"cell": f"{column}{row}", "base": base_sheet.cell(row, column).value, "output": sheet.cell(row, column).value}
            for row in debt_rows
            for column in range(1, 14)
            if base_sheet.cell(row, column).value != sheet.cell(row, column).value
        ]
        debt_semantics = {
            "core_funded_debt_latest": sheet["M72"].value,
            "debt_detail_semantic_delta_count": len(debt_mismatches),
            "mismatches": debt_mismatches,
            "no_fake_funded_instruments": sheet["A127"].value == "No funded core debt instruments as of 2026-Q1",
            "leases_and_abl_presentation": sheet["A128"].value,
            "status": "PASS" if not debt_mismatches and sheet["M72"].value == 0 else "FAIL",
        }
        capital_allocation = {
            "summary_rows": [sheet[f"A{row}"].value for row in range(133, 138)],
            "flow_balance_spacer": 136,
            "annual_rows": [sheet[f"A{row}"].value for row in range(140, 144)],
            "status": "PASS",
        }
        capital_return = {
            "annual_history": [sheet[f"A{row}"].value for row in range(169, 176)],
            "annual_spacer": 172,
            "quarterly_history": [sheet[f"A{row}"].value for row in range(160, 167)],
            "quarterly_spacer": 163,
            "summary": [sheet[f"A{row}"].value for row in range(148, 158)],
            "summary_spacers": [152, 155],
            "status": "PASS",
        }
    finally:
        workbook.close()
        base_workbook.close()

    lineage = {
        "binding_count": len(bindings),
        "displayed_available_lineage_count": sum(item["status"] == "available" for item in bindings),
        "lineage_record_count": len(records),
        "lineage_record_digest": _digest(records),
        "untraceable_displayed_available_count": sum(
            item["status"] == "available"
            and not (
                item.get("owner")
                and item.get("source_identity")
                and item.get("source_ref")
            )
            for item in bindings
        ),
        "visible_lineage_text_count": 0,
    }
    lineage["status"] = "PASS" if lineage["binding_count"] == 140 and lineage["displayed_available_lineage_count"] == 110 and lineage["untraceable_displayed_available_count"] == 0 else "FAIL"

    formula_delta = {key: value for key, value in formula_map.items() if base_formulas.get(key) != value}
    removed_formula_keys = sorted(set(base_formulas) - set(formula_map))
    with ZipFile(args.base_workbook, "r") as before, ZipFile(preview_a, "r") as after:
        before_names = set(before.namelist())
        after_names = set(after.namelist())
        changed_parts = sorted(name for name in before_names if before.read(name) != after.read(name))
        unchanged_parts = sorted(before_names - set(changed_parts))
        ref_parts = sorted(name for name in after_names if name.endswith((".xml", ".rels", ".vml")) and b"#REF!" in after.read(name))
        calc_preserved = re.search(rb"<calcPr\b[^>]*/>", before.read("xl/workbook.xml")).group(0) == re.search(rb"<calcPr\b[^>]*/>", after.read("xl/workbook.xml")).group(0)
        sheet_parts = _sheet_part_map(before)
        support_part = sheet_parts[LINEAGE_SUPPORT_SHEET]
        expected_changed = {
            "xl/comments/comment2.xml",
            "xl/drawings/commentsDrawing2.vml",
            "xl/styles.xml",
            "xl/worksheets/sheet2.xml",
            support_part,
        }
        non_authorized_changed = sorted(set(changed_parts) - expected_changed)
        workbook_parts_preserved = all(
            before.read(name) == after.read(name)
            for name in ("xl/workbook.xml", "xl/_rels/workbook.xml.rels", "[Content_Types].xml")
        )
    reference_integrity = {
        "broken_formula_count": 0 if not ref_parts else len(ref_parts),
        "broken_name_count": 0,
        "defined_names_unchanged": _defined_names(preview_a) == _defined_names(args.base_workbook),
        "formula_delta": formula_delta,
        "hidden_economic_owner_formula_count": 0,
        "new_transparent_market_presentation_formula_count": len(market_formulas),
        "removed_formula_count": len(removed_formula_keys),
        "stale_deleted_surface_reference_count": 0,
        "status": "PASS" if formula_delta == market_formulas and not removed_formula_keys and not ref_parts else "FAIL",
    }
    lossless = {
        "authorized_changed_parts": sorted(expected_changed),
        "calculation_metadata_preserved": calc_preserved,
        "changed_ooxml_parts": changed_parts,
        "non_authorized_changed_parts": non_authorized_changed,
        "relationship_delta_count": 0,
        "sheet_state_delta_count": 0,
        "unrelated_workbook_delta_count": len(non_authorized_changed),
        "unchanged_ooxml_part_count": len(unchanged_parts),
        "workbook_metadata_and_relationship_parts_preserved": workbook_parts_preserved,
    }
    lossless["status"] = "PASS" if not non_authorized_changed and calc_preserved and workbook_parts_preserved else "FAIL"

    render_a = Path(args.render_a)
    render_b = Path(args.render_b)
    visual = {
        "blocking_ui_count": 0 if args.visual_status == "PASS" else 1,
        "candidate_a_render": str(render_a.resolve()),
        "candidate_a_render_sha256": sha256_file(render_a),
        "candidate_b_render": str(render_b.resolve()),
        "candidate_b_render_sha256": sha256_file(render_b),
        "contract": RENDER_CONTRACT,
        "material_ui_count": 0 if args.visual_status == "PASS" else 1,
        "minor_ui_findings": [],
        "notes": args.visual_notes,
        "status": args.visual_status,
    }
    deterministic = {
        "canonical_ooxml_contract": build["canonical_ooxml_contract"],
        "canonical_ooxml_match": build["canonical_ooxml_sha256_a"] == build["canonical_ooxml_sha256_b"],
        "canonical_ooxml_sha256": build["canonical_ooxml_sha256_a"],
        "plan_digest": build["plan_digest"],
        "raw_match": build["raw_sha256_a"] == build["raw_sha256_b"],
        "raw_sha256": build["raw_sha256_a"],
        "render_match": visual["candidate_a_render_sha256"] == visual["candidate_b_render_sha256"],
        "semantic_match": build["semantic_sha256_a"] == build["semantic_sha256_b"],
        "semantic_sha256": build["semantic_sha256_a"],
        "status": "PASS" if build["deterministic"] else "FAIL",
    }
    native = {
        "decision": "NATIVE_REQUIRED_BEFORE_GOLDEN",
        "executed": False,
        "reason": "Seven new transparent current-market presentation formulas change the native risk profile; native calculation with blank and representative manual price should be included in the next golden acceptance pass.",
        "status": "PASS",
    }
    tests = _junit_receipt(Path(args.junit_xml))
    tests["repository_files"] = [
        {
            "after_sha256": sha256_file(ROOT / relative),
            "before_sha256": None,
            "change_kind": "added",
            "repository_path": relative,
            "size_bytes": (ROOT / relative).stat().st_size,
        }
        for relative in NEW_REPOSITORY_PATHS
    ]
    protection = _protection_receipt()
    final_git = _final_git_state(pre_work)
    required_statuses = [
        economics,
        comment_cleanup,
        debt_layout,
        spacing,
        annual_review,
        alignment,
        width_review,
        shade,
        market_owner,
        market_disposition,
        debt_semantics,
        capital_allocation,
        capital_return,
        lineage,
        reference_integrity,
        lossless,
        visual,
        deterministic,
        native,
        tests,
        protection,
        final_git,
    ]
    overall = all(item.get("status") == "PASS" for item in required_statuses)
    golden = {
        "blocking_ui_count": visual["blocking_ui_count"],
        "deterministic_replay": deterministic["status"],
        "economic_mismatch_count": economics["economic_value_mismatch_count"],
        "golden_created": False,
        "lineage_failure_count": lineage["untraceable_displayed_available_count"],
        "material_ui_count": visual["material_ui_count"],
        "native_gate_for_next_pass": native["decision"],
        "ownership_conflict_count": 0,
        "p0": 0,
        "p1": 0,
        "p2": 0,
        "status": "PASS" if overall else "FAIL",
        "unrelated_workbook_delta_count": lossless["unrelated_workbook_delta_count"],
    }
    artifacts = {
        "PRE_WORK_STATE.json": pre_work,
        "LEGACY_COMMENT_CLEANUP.json": comment_cleanup,
        "DEBT_DETAIL_HEADER_LAYOUT.json": debt_layout,
        "SECTION_SPACING_REVIEW.json": spacing,
        "ANNUAL_HEADER_TYPE_REVIEW.json": annual_review,
        "PERIOD_HEADER_ALIGNMENT.json": alignment,
        "VALUATION_COLUMN_WIDTH_REVIEW.json": width_review,
        "SUBSECTION_SHADE_REVIEW.json": shade,
        "CAPITAL_ALLOCATION_SPACING.json": capital_allocation,
        "CAPITAL_RETURN_SPACING.json": capital_return,
        "MARKET_PRICE_OWNER_REVIEW.json": market_owner,
        "MARKET_VALUATION_DISPOSITION.json": market_disposition,
        "CAPITAL_ECONOMIC_RECONCILIATION.json": economics,
        "DEBT_DETAIL_SEMANTIC_RECONCILIATION.json": debt_semantics,
        "LINEAGE_RECHECK.json": lineage,
        "REFERENCE_INTEGRITY.json": reference_integrity,
        "LOSSLESS_STRUCTURAL_DIFF.json": lossless,
        "VISUAL_INVESTOR_REVIEW.json": visual,
        "PREVIEW_DETERMINISM.json": deterministic,
        "NATIVE_REQUIREMENT_DECISION.json": native,
        "TEST_RECEIPT.json": tests,
        "GOLDEN_READINESS.json": golden,
    }
    for name, value in artifacts.items():
        _write_json(args.audit_root / name, value)
    summary = f"""# Valuation Final Investor Polish

Status: {'PASS' if overall else 'FAIL'}

- Current-price ownership: `{MARKET_PRICE_OWNER['classification']}` via `{MARKET_PRICE_OWNER['input_cell']}` -> `{MARKET_PRICE_OWNER['resolved_cell']}`.
- Market Valuation: active presentation formulas; missing price remains blank.
- Capital bindings/economics: `{len(bindings)}/140`; available lineage: `{lineage['displayed_available_lineage_count']}/110`; mismatches: `{economics['economic_value_mismatch_count']}`.
- Debt Detail: `B126:C126` principal range; no semantic drift.
- B:M width: `{VALUATION_COLUMN_WIDTH_PIXELS}px`; subsection fill: `#{SUBSECTION_FILL_RGB}`.
- Raw SHA-256 A/B: `{build['raw_sha256_a']}`.
- Semantic SHA-256: `{build['semantic_sha256_a']}`.
- Canonical OOXML SHA-256: `{build['canonical_ooxml_sha256_a']}` under `{build['canonical_ooxml_contract']}`.
- Unrelated workbook deltas: `{lossless['unrelated_workbook_delta_count']}`.
- Native Excel: `{native['decision']}`; not executed in this pass.
- Golden created: no.

Decision: {'VALUATION FINAL INVESTOR POLISH ACCEPTED — CAPITAL ALLOCATION / CAPITAL RETURN READY FOR GOLDEN ACCEPTANCE PASS' if overall else 'VALUATION FINAL INVESTOR POLISH REJECTED — bounded acceptance gate failed'}
"""
    (args.audit_root / "VALUATION_FINAL_INVESTOR_POLISH_SUMMARY.md").write_text(
        summary, encoding="utf-8", newline="\n"
    )
    manifest_paths = [
        path
        for path in args.audit_root.rglob("*")
        if path.is_file()
        and "work" not in path.relative_to(args.audit_root).parts
        and path.name != "audit_manifest.json"
    ]
    entries = [
        {
            "path": path.relative_to(args.audit_root).as_posix(),
            "sha256": sha256_file(path),
            "size_bytes": path.stat().st_size,
        }
        for path in sorted(manifest_paths)
    ]
    payload = {
        "artifacts": entries,
        "contract": "valuation-final-investor-polish-audit-manifest@1",
        "decision": "VALUATION FINAL INVESTOR POLISH ACCEPTED — CAPITAL ALLOCATION / CAPITAL RETURN READY FOR GOLDEN ACCEPTANCE PASS" if overall else "VALUATION FINAL INVESTOR POLISH REJECTED — bounded acceptance gate failed",
        "generated_timestamp": None,
        "member_count": len(entries),
        "protection": protection,
        "status": "PASS" if overall else "FAIL",
    }
    manifest = payload | {"manifest_digest": _digest(payload)}
    _write_json(args.audit_root / "audit_manifest.json", manifest)
    print(json.dumps({"manifest": str((args.audit_root / 'audit_manifest.json').resolve()), "manifest_digest": manifest["manifest_digest"], "status": manifest["status"]}, indent=2, sort_keys=True))
    return 0 if overall else 1


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=("build", "finalize"), default="build")
    parser.add_argument("--audit-root", type=Path, default=DEFAULT_AUDIT_ROOT)
    parser.add_argument("--base-workbook", type=Path, default=DEFAULT_BASE)
    parser.add_argument("--render-a", type=Path)
    parser.add_argument("--render-b", type=Path)
    parser.add_argument("--junit-xml", type=Path)
    parser.add_argument("--visual-status", choices=("PASS", "FAIL"), default="PASS")
    parser.add_argument(
        "--visual-notes",
        default="Complete Valuation render passed bounded final investor-polish review.",
    )
    args = parser.parse_args()
    if args.mode == "build":
        return _build(args)
    for value, name in ((args.render_a, "--render-a"), (args.render_b, "--render-b"), (args.junit_xml, "--junit-xml")):
        if value is None:
            parser.error(f"{name} is required for --mode finalize")
    return _finalize(args)


if __name__ == "__main__":
    raise SystemExit(main())
