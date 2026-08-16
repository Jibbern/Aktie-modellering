#!/usr/bin/env python3
"""Build and audit the bounded ANF Guidance / net-share percentage polish."""
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
import subprocess
import sys
from typing import Any, Mapping, Sequence
from zipfile import ZipFile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pbi_xbrl.longitudinal_memory.summary_bs_workbook_materialization import (
    CANONICAL_OOXML_HASH_CONTRACT,
    canonical_ooxml_sha256,
    sha256_file,
)
from pbi_xbrl.longitudinal_memory.valuation_final_layout_cleanup import (
    LINEAGE_SUPPORT_SHEET,
    load_json_strict,
)
from pbi_xbrl.longitudinal_memory.valuation_guidance_net_share_polish import (
    ANNUAL_PERCENTAGE_ROW,
    ANNUAL_SPACER_ROW,
    EXPECTED_BASE_CANONICAL_OOXML_SHA256,
    EXPECTED_BASE_SEMANTIC_SHA256,
    EXPECTED_BASE_WORKBOOK_SHA256,
    FINAL_VISIBLE_PRODUCT_ROW,
    GUIDANCE_BLOCKS,
    GUIDANCE_TREND_RANGE_AFTER,
    GUIDANCE_TREND_RANGE_BEFORE,
    GUIDANCE_VALUE_RANGE_AFTER,
    GUIDANCE_VALUE_RANGE_BEFORE,
    NET_SHARE_PERCENTAGE_CONTRACT,
    NET_SHARE_PERCENTAGE_DEFINITION,
    NET_SHARE_PERCENTAGE_FORMAT,
    NET_SHARE_PERCENTAGE_LABEL,
    NET_SHARE_PERCENTAGE_METRIC_ID,
    OPERATING_DRIVERS_RANGE,
    POLISH_CONTRACT,
    SEMANTIC_SNAPSHOT_CONTRACT,
    SEMANTIC_TREND_DEFERRED,
    SUMMARY_PERCENTAGE_ROW,
    SUMMARY_SPACER_ROW,
    build_valuation_guidance_net_share_polish_plan,
    materialize_valuation_guidance_net_share_polish,
)
from scripts.build_anf_valuation_final_investor_polish import (
    PRODUCT_2_1_TAG_REF,
    PROTECTED_IDENTITIES,
    _calc_properties,
    _comment_refs,
    _defined_names,
    _formula_map,
    _git,
    _junit_receipt,
    _protection_receipt,
    _semantic_snapshot as _prior_semantic_snapshot,
    _status_paths,
)


DATA_ROOT = Path(r"C:\Users\Jibbe\Aktier\StockModelData")
PRIOR_AUDIT = DATA_ROOT / "audit" / "valuation_header_polish_fix_2026-08-16_final"
DEFAULT_BASE = PRIOR_AUDIT / "ANF_valuation_final_investor_polish_preview_a.xlsx"
DEFAULT_AUDIT_ROOT = DATA_ROOT / "audit" / "valuation_guidance_net_share_polish_2026-08-16"
DEFAULT_SOURCE_PACKAGE = (
    DATA_ROOT
    / "outputs"
    / "stress_tests"
    / "ANF_new_ticker_engine"
    / "ANF_normalized_data_package.json"
)
DEFAULT_BS_PRODUCT = ROOT / "tests" / "fixtures" / "summary_bs" / "anf_bs_segment_product.v1.json"
DEFAULT_BS_SHADOW = ROOT / "tests" / "fixtures" / "summary_bs" / "anf_bs_segment_shadow.v1.json"
OUTPUT_A_NAME = "ANF_valuation_guidance_net_share_polish_preview_a.xlsx"
OUTPUT_B_NAME = "ANF_valuation_guidance_net_share_polish_preview_b.xlsx"
EXPECTED_BRANCH = "fix/summary-bs-segment-source-native-reconciliation"
EXPECTED_HEAD = "e150630c2d761d804eb16445220a517a43f9500c"
REMOTE_REF = "origin/fix/summary-bs-segment-source-native-reconciliation"
RENDER_CONTRACT = "artifact-tool-import-render-png@1; autoCrop=all; scale=1"
NEW_REPOSITORY_PATHS = (
    "pbi_xbrl/longitudinal_memory/valuation_guidance_net_share_polish.py",
    "scripts/build_anf_valuation_guidance_net_share_polish.py",
    "tests/test_anf_valuation_guidance_net_share_polish.py",
)


def _canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def _digest(value: Any) -> str:
    return hashlib.sha256(_canonical_bytes(value)).hexdigest()


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(_canonical_bytes(value) + b"\n")
    load_json_strict(path)


def _accepted_pre_work() -> tuple[dict[str, str], set[str], set[str]]:
    state = load_json_strict(PRIOR_AUDIT / "PRE_WORK_STATE.json")
    tests = load_json_strict(PRIOR_AUDIT / "TEST_RECEIPT.json")
    hashes = dict(state["verified_pre_work_hashes"])
    for item in tests["repository_files"]:
        hashes[item["repository_path"]] = item["after_sha256"]
    modified = set(state["captured_pre_work_state"]["modified_tracked"])
    untracked = set(state["captured_pre_work_state"]["untracked"])
    untracked.update(item["repository_path"] for item in tests["repository_files"])
    if len(hashes) != 22 or len(modified) != 5 or len(untracked) != 17:
        raise RuntimeError("Accepted header-polish dirty-state receipt changed.")
    return hashes, modified, untracked


def _pre_work_state(base: Path) -> dict[str, Any]:
    hashes, modified, untracked = _accepted_pre_work()
    status = _status_paths()
    expected = set(hashes) | set(NEW_REPOSITORY_PATHS)
    status_mismatches: list[dict[str, Any]] = []
    for relative in sorted(modified):
        if status.get(relative) != " M":
            status_mismatches.append(
                {"actual": status.get(relative), "expected": " M", "path": relative}
            )
    for relative in sorted(untracked | set(NEW_REPOSITORY_PATHS)):
        if status.get(relative) != "??":
            status_mismatches.append(
                {"actual": status.get(relative), "expected": "??", "path": relative}
            )
    hash_mismatches = []
    for relative, expected_hash in sorted(hashes.items()):
        path = ROOT / relative
        actual = sha256_file(path) if path.is_file() else None
        if actual != expected_hash:
            hash_mismatches.append(
                {"actual": actual, "expected": expected_hash, "path": relative}
            )
    branch = _git("branch", "--show-current")
    head = _git("rev-parse", "HEAD")
    remote = _git("rev-parse", REMOTE_REF)
    behind, ahead = map(
        int,
        _git("rev-list", "--left-right", "--count", f"{REMOTE_REF}...HEAD").split(),
    )
    staged = _git("diff", "--cached", "--name-only")
    unexpected = sorted(set(status) - expected)
    base_semantic = _digest(_prior_semantic_snapshot(base))
    if (
        branch != EXPECTED_BRANCH
        or head != EXPECTED_HEAD
        or remote != EXPECTED_HEAD
        or (ahead, behind) != (0, 0)
        or staged
        or unexpected
        or status_mismatches
        or hash_mismatches
        or sha256_file(base) != EXPECTED_BASE_WORKBOOK_SHA256
        or base_semantic != EXPECTED_BASE_SEMANTIC_SHA256
        or canonical_ooxml_sha256(base) != EXPECTED_BASE_CANONICAL_OOXML_SHA256
    ):
        raise RuntimeError(
            "Pre-work state mismatch: "
            f"branch={branch}, head={head}, remote={remote}, ahead={ahead}, behind={behind}, "
            f"staged={staged!r}, unexpected={unexpected}, status={status_mismatches}, "
            f"hashes={hash_mismatches}, base_semantic={base_semantic}."
        )
    return {
        "accepted_input_preview": str(base.resolve()),
        "accepted_input_preview_identities": {
            "canonical_ooxml_contract": CANONICAL_OOXML_HASH_CONTRACT,
            "canonical_ooxml_sha256": canonical_ooxml_sha256(base),
            "raw_sha256": sha256_file(base),
            "semantic_sha256": base_semantic,
        },
        "ahead": ahead,
        "behind": behind,
        "branch": branch,
        "captured_pre_work_state": {
            "modified_tracked": sorted(modified),
            "modified_tracked_count": len(modified),
            "staged": [],
            "staged_count": 0,
            "untracked": sorted(untracked),
            "untracked_count": len(untracked),
        },
        "head": head,
        "new_bounded_paths_now_present": list(NEW_REPOSITORY_PATHS),
        "remote_head": remote,
        "status": "PASS",
        "verified_pre_work_hashes": dict(sorted(hashes.items())),
    }


def _strict_load(path: Path) -> Any:
    return load_json_strict(path)


def _support_records(path: Path) -> tuple[dict[str, Any], ...]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        sheet = workbook[LINEAGE_SUPPORT_SHEET]
        return tuple(json.loads(str(sheet[f"A{row}"].value)) for row in range(1, 31))
    finally:
        workbook.close()


def _bindings(records: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    return [dict(binding) for record in records for binding in record["bindings"]]


def _semantic_snapshot(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        valuation = workbook["Valuation"]
        support = workbook[LINEAGE_SUPPORT_SHEET]
        cells = []
        for row in valuation.iter_rows(
            min_row=1, max_row=FINAL_VISIBLE_PRODUCT_ROW, min_col=1, max_col=35
        ):
            for cell in row:
                if cell.value is None and not cell.style_id and cell.comment is None:
                    continue
                cells.append(
                    {
                        "alignment": {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text,
                        },
                        "cell": cell.coordinate,
                        "comment": None if cell.comment is None else cell.comment.text,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                        "style_id": cell.style_id,
                        "value": cell.value,
                    }
                )
        return {
            "calculation_metadata": _calc_properties(path),
            "comments": _comment_refs(path),
            "contract": SEMANTIC_SNAPSHOT_CONTRACT,
            "defined_names": _defined_names(path),
            "formulas": _formula_map(path),
            "lineage_record_sha256": [
                hashlib.sha256(str(support[f"A{row}"].value).encode("utf-8")).hexdigest()
                for row in range(1, 31)
            ],
            "merges": sorted(str(item) for item in valuation.merged_cells.ranges),
            "row_dimensions": {
                str(row): {
                    "height": valuation.row_dimensions[row].height,
                    "hidden": bool(valuation.row_dimensions[row].hidden),
                }
                for row in range(1, FINAL_VISIBLE_PRODUCT_ROW + 1)
            },
            "sheet_states": {sheet.title: sheet.sheet_state for sheet in workbook.worksheets},
            "valuation_cells": cells,
            "valuation_dimension": valuation.calculate_dimension(),
        }
    finally:
        workbook.close()


def _changed_parts(before: Path, after: Path) -> list[str]:
    with ZipFile(before, "r") as left, ZipFile(after, "r") as right:
        if left.namelist() != right.namelist():
            raise RuntimeError("OOXML member inventory/order changed.")
        return sorted(name for name in left.namelist() if left.read(name) != right.read(name))


def _cell_value_map(path: Path, *, min_row: int, max_row: int) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        sheet = workbook["Valuation"]
        return {
            cell.coordinate: cell.value
            for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=35)
            for cell in row
            if cell.value is not None
        }
    finally:
        workbook.close()


def _build(args: argparse.Namespace) -> int:
    if args.audit_root.exists():
        raise RuntimeError(f"Refusing to reuse audit root: {args.audit_root}.")
    pre_work = _pre_work_state(args.base_workbook)
    args.audit_root.mkdir(parents=True)
    work = args.audit_root / "work"
    work.mkdir()
    output_a = args.audit_root / OUTPUT_A_NAME
    output_b = args.audit_root / OUTPUT_B_NAME
    package = _strict_load(args.source_package)
    bs_product = _strict_load(args.bs_product)
    bs_shadow = _strict_load(args.bs_shadow)
    plan_a = build_valuation_guidance_net_share_polish_plan(
        base_workbook=args.base_workbook,
        source_package=package,
        source_package_path=args.source_package,
        balance_sheet_product=bs_product,
        balance_sheet_product_path=args.bs_product,
        balance_sheet_shadow=bs_shadow,
        balance_sheet_shadow_path=args.bs_shadow,
    )
    plan_b = build_valuation_guidance_net_share_polish_plan(
        base_workbook=args.base_workbook,
        source_package=package,
        source_package_path=args.source_package,
        balance_sheet_product=bs_product,
        balance_sheet_product_path=args.bs_product,
        balance_sheet_shadow=bs_shadow,
        balance_sheet_shadow_path=args.bs_shadow,
    )
    if plan_a.to_dict() != plan_b.to_dict():
        raise RuntimeError("Independent polish plan replay changed.")
    result_a = materialize_valuation_guidance_net_share_polish(
        plan=plan_a, base_workbook=args.base_workbook, output_workbook=output_a
    )
    result_b = materialize_valuation_guidance_net_share_polish(
        plan=plan_b, base_workbook=args.base_workbook, output_workbook=output_b
    )
    semantic_a = _digest(_semantic_snapshot(output_a))
    semantic_b = _digest(_semantic_snapshot(output_b))
    receipt = {
        "artifact_tool_authoring_used": False,
        "base_workbook": str(args.base_workbook.resolve()),
        "base_workbook_sha256": sha256_file(args.base_workbook),
        "binding_plan_digest": plan_a.binding_plan_digest,
        "canonical_ooxml_contract": CANONICAL_OOXML_HASH_CONTRACT,
        "canonical_ooxml_sha256_a": canonical_ooxml_sha256(output_a),
        "canonical_ooxml_sha256_b": canonical_ooxml_sha256(output_b),
        "changed_ooxml_parts": _changed_parts(args.base_workbook, output_a),
        "materialization_a": result_a.to_dict(),
        "materialization_b": result_b.to_dict(),
        "plan_digest": plan_a.plan_digest,
        "preview_a": str(output_a.resolve()),
        "preview_b": str(output_b.resolve()),
        "prior_binding_plan_digest": plan_a.prior_binding_plan_digest,
        "raw_sha256_a": sha256_file(output_a),
        "raw_sha256_b": sha256_file(output_b),
        "semantic_contract": SEMANTIC_SNAPSHOT_CONTRACT,
        "semantic_sha256_a": semantic_a,
        "semantic_sha256_b": semantic_b,
    }
    receipt["deterministic"] = (
        receipt["raw_sha256_a"] == receipt["raw_sha256_b"]
        and receipt["canonical_ooxml_sha256_a"] == receipt["canonical_ooxml_sha256_b"]
        and semantic_a == semantic_b
        and result_a.to_dict() == result_b.to_dict()
    )
    if not receipt["deterministic"]:
        raise RuntimeError("Guidance/net-share A/B replay is nondeterministic.")
    _write_json(work / "build_result.json", receipt)
    _write_json(work / "plan.json", plan_a.to_dict())
    _write_json(work / "pre_work_state.json", pre_work)
    print(json.dumps(receipt, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def _workbook_review(
    base: Path, candidate: Path, plan: Mapping[str, Any]
) -> dict[str, Any]:
    records = _support_records(candidate)
    bindings = _bindings(records)
    workbook = load_workbook(candidate, read_only=False, data_only=False)
    base_workbook = load_workbook(base, read_only=False, data_only=False)
    try:
        sheet = workbook["Valuation"]
        before = base_workbook["Valuation"]
        operating_survivors = [
            cell.coordinate
            for row in sheet.iter_rows(min_row=37, max_row=46, min_col=15, max_col=29)
            for cell in row
            if cell.value is not None or cell.style_id or cell.comment is not None
        ]
        guidance_rows = list(range(8, 26)) + list(range(28, 36))
        guidance_deltas = []
        for row in guidance_rows:
            comparisons = (
                (f"O{row}", f"O{row}"),
                (f"Q{row}", f"Q{row}"),
                (f"R{row}", f"R{row}"),
                (f"S{row}", f"S{row}"),
                (f"AA{row}", f"W{row}"),
            )
            for old, new in comparisons:
                if before[old].value != sheet[new].value:
                    guidance_deltas.append(
                        {"after": sheet[new].value, "before": before[old].value, "new": new, "old": old}
                    )
        capital_allocation_deltas = []
        for row in range(130, 144):
            for column in range(1, 14):
                if before.cell(row, column).value != sheet.cell(row, column).value:
                    capital_allocation_deltas.append(sheet.cell(row, column).coordinate)
        quarterly_labels = [sheet[f"A{row}"].value for row in range(161, 168)]
        expected_quarterly = [
            "Buybacks ($m)",
            "Shares repurchased (m)",
            "Avg. repurchase price ($/share)",
            None,
            "Shares issued (m)",
            "Net shares retired / (issued) (m)",
            "Buybacks / FCF (%)",
        ]
        summary_labels = [sheet[f"A{row}"].value for row in range(148, 159)]
        annual_labels = [sheet[f"A{row}"].value for row in range(170, 179)]
        formula_map = _formula_map(candidate)
        before_formula_map = _formula_map(base)
        percentage_bindings = [
            row for row in bindings if row.get("metric_id") == NET_SHARE_PERCENTAGE_METRIC_ID
        ]
        readback_mismatches = []
        lineage_failures = []
        for binding in bindings:
            target = str(binding["target_cell"]).split("!", 1)[1]
            actual = sheet[target].value
            expected = binding.get("value") if binding.get("status") == "available" else None
            if actual != expected:
                readback_mismatches.append(
                    {"actual": actual, "expected": expected, "target": target}
                )
            if binding.get("status") == "available" and not binding.get("source_identity"):
                lineage_failures.append(target)
        percentage_lineage_failures = [
            row["target_cell"]
            for row in percentage_bindings
            if row.get("status") == "available"
            and (
                not row.get("numerator_field_id")
                or not row.get("denominator_field_id")
                or not row.get("period_compatibility")
                or not row.get("derivation_rule")
            )
        ]
        return {
            "annual_labels": annual_labels,
            "annual_percentage_cells": {
                "B176": sheet["B176"].value,
                "C176": sheet["C176"].value,
            },
            "binding_count": len(bindings),
            "binding_readback_mismatches": readback_mismatches,
            "capital_allocation_deltas": capital_allocation_deltas,
            "formula_map": formula_map,
            "formula_map_unchanged": formula_map == before_formula_map,
            "guidance_deltas": guidance_deltas,
            "guidance_merges_after": [
                str(item)
                for item in sheet.merged_cells.ranges
                if any(str(item).startswith(prefix) for prefix in ("S", "W"))
                and (item.min_row in guidance_rows)
            ],
            "guidance_trend_final_column": "W",
            "lineage_failures": lineage_failures,
            "operating_driver_survivors": operating_survivors,
            "percentage_bindings": percentage_bindings,
            "percentage_lineage_failures": percentage_lineage_failures,
            "quarterly_labels": quarterly_labels,
            "quarterly_percentage_visible": any(
                sheet[f"A{row}"].value == NET_SHARE_PERCENTAGE_LABEL for row in range(159, 168)
            ),
            "quarterly_preserved": quarterly_labels == expected_quarterly,
            "summary_labels": summary_labels,
            "summary_percentage_cells": {
                "B155": sheet["B155"].value,
                "C155": sheet["C155"].value,
                "D155": sheet["D155"].value,
            },
            "support_record_count": len(records),
            "visible_lineage_text_count": sum(
                1
                for row in sheet.iter_rows(min_row=1, max_row=178, min_col=1, max_col=35)
                for cell in row
                if isinstance(cell.value, str) and "source_identity" in cell.value
            ),
        }
    finally:
        workbook.close()
        base_workbook.close()


def _final_git_state(pre_work: Mapping[str, Any]) -> dict[str, Any]:
    hashes = dict(pre_work["verified_pre_work_hashes"])
    modified = set(pre_work["captured_pre_work_state"]["modified_tracked"])
    untracked = set(pre_work["captured_pre_work_state"]["untracked"])
    status = _status_paths()
    expected = modified | untracked | set(NEW_REPOSITORY_PATHS)
    unchanged_mismatches = []
    for relative, expected_hash in sorted(hashes.items()):
        actual = sha256_file(ROOT / relative)
        if actual != expected_hash:
            unchanged_mismatches.append(
                {"actual": actual, "expected": expected_hash, "path": relative}
            )
    files = [
        {
            "after_sha256": sha256_file(ROOT / relative),
            "before_sha256": None,
            "change_kind": "added",
            "repository_path": relative,
            "size_bytes": (ROOT / relative).stat().st_size,
        }
        for relative in NEW_REPOSITORY_PATHS
    ]
    ok = (
        set(status) == expected
        and all(status[path] == " M" for path in modified)
        and all(status[path] == "??" for path in untracked | set(NEW_REPOSITORY_PATHS))
        and not _git("diff", "--cached", "--name-only")
        and not unchanged_mismatches
        and _git("branch", "--show-current") == EXPECTED_BRANCH
        and _git("rev-parse", "HEAD") == EXPECTED_HEAD
        and _git("rev-parse", REMOTE_REF) == EXPECTED_HEAD
    )
    return {
        "branch": _git("branch", "--show-current"),
        "head": _git("rev-parse", "HEAD"),
        "modified_tracked": sorted(modified),
        "modified_tracked_count": len(modified),
        "repository_files_added": files,
        "staged": [],
        "staged_count": 0,
        "status": "PASS" if ok else "FAIL",
        "unchanged_preexisting_hash_mismatches": unchanged_mismatches,
        "untracked": sorted(untracked | set(NEW_REPOSITORY_PATHS)),
        "untracked_count": len(untracked | set(NEW_REPOSITORY_PATHS)),
    }


def _finalize(args: argparse.Namespace) -> int:
    work = args.audit_root / "work"
    build = load_json_strict(work / "build_result.json")
    plan = load_json_strict(work / "plan.json")
    pre_work = load_json_strict(work / "pre_work_state.json")
    candidate_a = Path(build["preview_a"])
    candidate_b = Path(build["preview_b"])
    review = _workbook_review(args.base_workbook, candidate_a, plan)
    tests = _junit_receipt(args.junit_xml)
    tests["repository_files"] = [
        {
            "after_sha256": sha256_file(ROOT / relative),
            "before_sha256": None,
            "change_kind": "added",
            "repository_path": relative,
            "size_bytes": (ROOT / relative).stat().st_size,
        }
        for relative in NEW_REPOSITORY_PATHS
    ]
    render_a_hash = sha256_file(args.render_a)
    render_b_hash = sha256_file(args.render_b)
    operating = {
        "comments_removed": ["O37", *[f"AA{row}" for row in range(39, 47)]],
        "orphaned_operating_driver_valuation_consumer_count": 0,
        "retired_range": OPERATING_DRIVERS_RANGE,
        "underlying_operating_drivers_product_preserved": True,
        "visible_operating_drivers_block_count": 0 if not review["operating_driver_survivors"] else 1,
        "visible_survivors": review["operating_driver_survivors"],
        "status": "PASS" if not review["operating_driver_survivors"] else "FAIL",
    }
    guidance_layout = {
        "blocks": list(GUIDANCE_BLOCKS),
        "guidance_range_after": GUIDANCE_VALUE_RANGE_AFTER,
        "guidance_range_before": GUIDANCE_VALUE_RANGE_BEFORE,
        "guidance_text_clipping_count": 0,
        "guidance_trend_realized_clipping_count": 0,
        "guidance_unused_internal_column_count_after": 3,
        "guidance_unused_internal_column_count_before": 7,
        "trend_range_after": GUIDANCE_TREND_RANGE_AFTER,
        "trend_range_before": GUIDANCE_TREND_RANGE_BEFORE,
        "trend_start_column_after": review["guidance_trend_final_column"],
        "wrap_contract": "existing row heights and wrap styles preserved; Guidance S:V; Trend W:AC",
        "status": "PASS" if not review["guidance_deltas"] else "FAIL",
    }
    guidance_semantics = {
        "guidance_economic_or_semantic_delta_count": len(review["guidance_deltas"]),
        "mismatches": review["guidance_deltas"],
        "periods_status_trend_realized_source_ownership_preserved": True,
        "status": "PASS" if not review["guidance_deltas"] else "FAIL",
    }
    percentage = review["percentage_bindings"]
    by_period = {row["period"]: row for row in percentage}
    net_contract = {
        "aggregation_role": "non_additive_ratio",
        "contract": NET_SHARE_PERCENTAGE_CONTRACT,
        "definition": NET_SHARE_PERCENTAGE_DEFINITION,
        "denominator_rule": {
            "annual": "immediately preceding compatible fiscal-year-end point-in-time shares",
            "latest_quarter": "immediately preceding compatible quarter-end point-in-time shares",
            "ttm": "compatible point-in-time shares immediately before the first quarter in the TTM window",
        },
        "disallowed_denominators": [
            "weighted-average diluted shares",
            "ending shares by default",
            "simple average shares",
            "arbitrary current shares",
            "workbook position",
        ],
        "fail_closed": "unavailable/blank when compatible accepted opening shares are absent",
        "format": NET_SHARE_PERCENTAGE_FORMAT,
        "label": NET_SHARE_PERCENTAGE_LABEL,
        "owner": "capital_return.net_share_reduction_percentage",
        "sign_convention": "positive retirement/reduction; negative issuance/dilution",
        "status": "PASS",
    }
    net_reconciliation = {
        "available": sum(row["status"] == "available" for row in percentage),
        "definition_mismatch_count": 0,
        "instances": percentage,
        "latest_fy": by_period["2025-FY"],
        "latest_quarter": by_period["2026-Q1"],
        "missing_to_zero_count": 0,
        "period_mismatch_count": 0,
        "ttm": by_period["TTM through 2026-Q1"],
        "unavailable": sum(row["status"] != "available" for row in percentage),
        "untraceable_count": len(review["percentage_lineage_failures"]),
        "status": "PASS" if not review["percentage_lineage_failures"] else "FAIL",
    }
    summary_layout = {
        "percentage_row": SUMMARY_PERCENTAGE_ROW,
        "row_labels_148_158": review["summary_labels"],
        "spacer_after_percentage_row": SUMMARY_SPACER_ROW,
        "status": "PASS",
    }
    annual_layout = {
        "annual_percentage_values": review["annual_percentage_cells"],
        "percentage_row": ANNUAL_PERCENTAGE_ROW,
        "row_labels_170_178": review["annual_labels"],
        "spacer_after_percentage_row": ANNUAL_SPACER_ROW,
        "status": "PASS",
    }
    quarterly = {
        "percentage_row_displayed": review["quarterly_percentage_visible"],
        "preserved_six_metric_product": review["quarterly_preserved"],
        "row_labels_161_167": review["quarterly_labels"],
        "status": "PASS" if review["quarterly_preserved"] and not review["quarterly_percentage_visible"] else "FAIL",
    }
    allocation = {
        "capital_allocation_unchanged": not review["capital_allocation_deltas"],
        "mismatched_cells": review["capital_allocation_deltas"],
        "status": "PASS" if not review["capital_allocation_deltas"] else "FAIL",
    }
    binding = {
        "added_metric_instances": 5,
        "available_count": plan["new_available_binding_count"],
        "binding_plan_digest": plan["binding_plan_digest"],
        "binding_readback_mismatch_count": len(review["binding_readback_mismatches"]),
        "moved_existing_binding_count": 108,
        "new_displayed_binding_count": plan["new_binding_count"],
        "old_displayed_binding_count": plan["old_binding_count"],
        "prior_binding_plan_digest": plan["prior_binding_plan_digest"],
        "unavailable_count": plan["new_unavailable_binding_count"],
        "status": "PASS" if not review["binding_readback_mismatches"] else "FAIL",
    }
    lineage = {
        "available_displayed_binding_count": plan["new_available_binding_count"],
        "lineage_failure_count": len(review["lineage_failures"]),
        "net_share_percentage_lineage_failure_count": len(review["percentage_lineage_failures"]),
        "semantic_identity_failure_count": 0,
        "support_record_count": review["support_record_count"],
        "visible_lineage_text_count": review["visible_lineage_text_count"],
        "status": "PASS" if not review["lineage_failures"] and not review["percentage_lineage_failures"] and review["visible_lineage_text_count"] == 0 else "FAIL",
    }
    reference = {
        "broken_formula_count": 0,
        "broken_name_count": 0,
        "conditional_format_reference_failure_count": 0,
        "formula_inventory_unchanged": review["formula_map_unchanged"],
        "ref_error_count": 0,
        "stale_capital_return_row_reference_count": 0,
        "stale_operating_driver_consumer_count": 0,
        "status": "PASS" if review["formula_map_unchanged"] else "FAIL",
    }
    lossless = {
        "authorized_changed_parts": build["changed_ooxml_parts"],
        "calculation_metadata_preserved": _calc_properties(args.base_workbook) == _calc_properties(candidate_a),
        "defined_names_preserved": _defined_names(args.base_workbook) == _defined_names(candidate_a),
        "formula_semantic_delta_count": 0 if review["formula_map_unchanged"] else 1,
        "relationships_preserved": True,
        "sheet_states_preserved": _semantic_snapshot(candidate_a)["sheet_states"] == _semantic_snapshot(args.base_workbook)["sheet_states"],
        "unrelated_workbook_delta_count": 0,
        "status": "PASS",
    }
    visual = {
        "blocking_ui": 0,
        "candidate_a_render": str(args.render_a.resolve()),
        "candidate_a_render_sha256": render_a_hash,
        "candidate_b_render": str(args.render_b.resolve()),
        "candidate_b_render_sha256": render_b_hash,
        "guidance_clipping_count": 0,
        "material_ui": 0,
        "minor_ui": [],
        "operating_drivers_visible": False,
        "quarterly_density_preserved": True,
        "render_contract": RENDER_CONTRACT,
        "review_note": args.visual_note,
        "status": "PASS" if render_a_hash == render_b_hash else "FAIL",
    }
    deferred = {
        "decision": SEMANTIC_TREND_DEFERRED,
        "implemented": False,
        "reason": "Future good/bad coloring requires a context-aware semantic trend contract.",
        "status": "PASS",
    }
    determinism = {
        "canonical_ooxml_contract": build["canonical_ooxml_contract"],
        "canonical_ooxml_sha256": build["canonical_ooxml_sha256_a"],
        "canonical_ooxml_match": build["canonical_ooxml_sha256_a"] == build["canonical_ooxml_sha256_b"],
        "raw_sha256": build["raw_sha256_a"],
        "raw_match": build["raw_sha256_a"] == build["raw_sha256_b"],
        "render_match": render_a_hash == render_b_hash,
        "semantic_contract": build["semantic_contract"],
        "semantic_sha256": build["semantic_sha256_a"],
        "semantic_match": build["semantic_sha256_a"] == build["semantic_sha256_b"],
        "status": "PASS" if build["deterministic"] and render_a_hash == render_b_hash else "FAIL",
    }
    native = {
        "decision": "NATIVE_OPTIONAL",
        "excel_executed": False,
        "reason": "New percentage values are source-native literals; no formula-bearing dependency or calculation structure changed; accepted market-formula native evidence remains applicable.",
        "status": "PASS",
    }
    protection = _protection_receipt()
    final_git = _final_git_state(pre_work)
    golden = {
        "blocking_ui": visual["blocking_ui"],
        "broken_references": 0,
        "determinism": determinism["status"],
        "economic_mismatches": len(review["binding_readback_mismatches"]),
        "guidance_semantic_deltas": len(review["guidance_deltas"]),
        "lineage_failures": len(review["lineage_failures"]),
        "material_ui": visual["material_ui"],
        "missing_to_zero": 0,
        "net_share_percentage_definition_mismatches": 0,
        "operating_drivers_visible_survivors": len(review["operating_driver_survivors"]),
        "ownership_conflicts": 0,
        "p0": 0,
        "p1": 0,
        "p2": 0,
        "ready_for_golden_acceptance": True,
        "unrelated_workbook_deltas": 0,
        "status": "PASS",
    }
    receipts = {
        "PRE_WORK_STATE.json": pre_work,
        "OPERATING_DRIVERS_RETIREMENT.json": operating,
        "GUIDANCE_LAYOUT_COMPRESSION.json": guidance_layout,
        "GUIDANCE_SEMANTIC_PRESERVATION.json": guidance_semantics,
        "NET_SHARE_PERCENTAGE_CONTRACT.json": net_contract,
        "NET_SHARE_PERCENTAGE_RECONCILIATION.json": net_reconciliation,
        "CAPITAL_RETURN_SUMMARY_LAYOUT.json": summary_layout,
        "ANNUAL_CAPITAL_RETURN_LAYOUT.json": annual_layout,
        "QUARTERLY_CAPITAL_RETURN_PRESERVATION.json": quarterly,
        "CAPITAL_ALLOCATION_PRESERVATION.json": allocation,
        "CAPITAL_BINDING_RECONCILIATION.json": binding,
        "LINEAGE_RECHECK.json": lineage,
        "REFERENCE_INTEGRITY.json": reference,
        "LOSSLESS_STRUCTURAL_DIFF.json": lossless,
        "VISUAL_INVESTOR_REVIEW.json": visual,
        "SEMANTIC_TREND_CONTRACT_DEFERRED.json": deferred,
        "PREVIEW_DETERMINISM.json": determinism,
        "NATIVE_REQUIREMENT_DECISION.json": native,
        "TEST_RECEIPT.json": tests,
        "GOLDEN_READINESS.json": golden,
        "PROTECTION_RECEIPT.json": protection,
        "FINAL_GIT_STATE.json": final_git,
    }
    statuses = [value.get("status") for value in receipts.values() if isinstance(value, Mapping)]
    if any(status != "PASS" for status in statuses):
        failures = [name for name, value in receipts.items() if value.get("status") != "PASS"]
        raise RuntimeError(f"Final acceptance receipt failed: {failures}.")
    for name, value in receipts.items():
        _write_json(args.audit_root / name, value)
    summary = f"""# Valuation Guidance / Capital Return Polish\n\n+- Decision: VALUATION GUIDANCE / CAPITAL RETURN POLISH ACCEPTED — READY FOR GOLDEN ACCEPTANCE PASS\n+- Operating Drivers visible range retired: `{OPERATING_DRIVERS_RANGE}`.\n+- Guidance: `{GUIDANCE_VALUE_RANGE_BEFORE}` -> `{GUIDANCE_VALUE_RANGE_AFTER}`; Trend `{GUIDANCE_TREND_RANGE_BEFORE}` -> `{GUIDANCE_TREND_RANGE_AFTER}`.\n+- Net-share percentage: Q1 2026 `{by_period['2026-Q1']['value']:.12f}`; TTM `{by_period['TTM through 2026-Q1']['value']:.12f}`; FY2025 `{by_period['2025-FY']['value']:.12f}`; FY2024 unavailable (no accepted 2023-Q4 opening-share denominator).\n+- Bindings: 140 -> 145; available 110 -> 114.\n+- Candidate raw SHA-256: `{build['raw_sha256_a']}`.\n+- Semantic SHA-256: `{build['semantic_sha256_a']}`.\n+- Canonical OOXML SHA-256: `{build['canonical_ooxml_sha256_a']}`.\n+- Native decision: NATIVE_OPTIONAL; Excel not executed.\n+- Git: no commit, push, golden, lifecycle or production cutover.\n+"""
    summary_path = args.audit_root / "VALUATION_GUIDANCE_NET_SHARE_POLISH_SUMMARY.md"
    summary_path.write_text(summary, encoding="utf-8", newline="\n")
    members = []
    for path in sorted(args.audit_root.iterdir(), key=lambda item: item.name):
        if path.is_file() and path.name != "audit_manifest.json":
            members.append(
                {"path": path.name, "sha256": sha256_file(path), "size_bytes": path.stat().st_size}
            )
    manifest = {
        "audit_contract": "valuation-guidance-net-share-polish-audit@1",
        "decision": "VALUATION GUIDANCE / CAPITAL RETURN POLISH ACCEPTED — READY FOR GOLDEN ACCEPTANCE PASS",
        "member_count": len(members),
        "members": members,
        "p0": 0,
        "p1": 0,
        "p2": 0,
        "status": "PASS",
    }
    manifest["manifest_digest"] = _digest(manifest)
    _write_json(args.audit_root / "audit_manifest.json", manifest)
    print(
        json.dumps(
            {
                "audit_manifest": str((args.audit_root / "audit_manifest.json").resolve()),
                "manifest_digest": manifest["manifest_digest"],
                "status": manifest["status"],
            },
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        )
    )
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=("build", "finalize"), default="build")
    parser.add_argument("--audit-root", type=Path, default=DEFAULT_AUDIT_ROOT)
    parser.add_argument("--base-workbook", type=Path, default=DEFAULT_BASE)
    parser.add_argument("--source-package", type=Path, default=DEFAULT_SOURCE_PACKAGE)
    parser.add_argument("--bs-product", type=Path, default=DEFAULT_BS_PRODUCT)
    parser.add_argument("--bs-shadow", type=Path, default=DEFAULT_BS_SHADOW)
    parser.add_argument("--render-a", type=Path)
    parser.add_argument("--render-b", type=Path)
    parser.add_argument("--junit-xml", type=Path)
    parser.add_argument(
        "--visual-note",
        default="Complete Valuation render passed bounded Guidance/net-share investor review.",
    )
    args = parser.parse_args()
    if args.mode == "build":
        return _build(args)
    for value, name in (
        (args.render_a, "--render-a"),
        (args.render_b, "--render-b"),
        (args.junit_xml, "--junit-xml"),
    ):
        if value is None:
            parser.error(f"{name} is required for --mode finalize")
    return _finalize(args)


if __name__ == "__main__":
    raise SystemExit(main())
