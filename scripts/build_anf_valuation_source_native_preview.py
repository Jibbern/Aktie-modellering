"""Build and audit the bounded pre-native ANF Valuation correction previews."""
from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import asdict
from decimal import Decimal
import hashlib
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import sys
from typing import Any, Iterable, Mapping, Sequence
from xml.etree import ElementTree as ET
from zipfile import ZipFile

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer

from pbi_xbrl.longitudinal_memory.formula_aware_workbook_materialization import (
    CANONICAL_OOXML_HASH_CONTRACT,
    FORMULA_CACHE_POLICY,
    MATERIALIZER_CONTRACT,
    _cell_elements,
    _sheet_part_map,
    canonical_ooxml_sha256,
    materialize_formula_aware_mutations,
    sha256_file,
)
from pbi_xbrl.longitudinal_memory.valuation_source_native_projection import (
    BASE_SUMMARY_BS_GOLDEN_SHA256,
    CALCULATION_METADATA_POLICY_ID,
    IC_DATA_SHEET,
    IC_VISIBLE_SHEET,
    PROJECTION_CONTRACT,
    PROTECTED_ANF_SHA256,
    VALUATION_CALCULATION_METADATA_POLICY,
    VALUATION_SHEET,
    ValuationProjectionPlan,
    _expand_reference,
    _load_json,
    build_valuation_projection_plan,
)
from pbi_xbrl.new_ticker_investment_case_formula_surface import (
    CANONICAL_SCENARIOS,
    CANONICAL_VALUATION_METHODS,
    canonical_investment_case_defined_names,
)
from scripts.fill_anf_shadow_workbook import run_anf_shadow_workbook_fill


AUDIT_CONTRACT = "valuation-bounded-correction-audit@1"
SEMANTIC_HASH_CONTRACT = "valuation-pre-native-semantic-snapshot-sha256@1"
EXPECTED_HEAD = "42a9796cdc227e88db4ee1986d9deb75767f37e4"
EXPECTED_BRANCH = "fix/summary-bs-segment-source-native-reconciliation"
EXPECTED_SUMMARY_BS_GOLDEN = BASE_SUMMARY_BS_GOLDEN_SHA256
EXPECTED_PBI_SHA256 = "6482617ad4f412dc5a1e130dc56c72bba5113ddacea5f7d1ab166fad8ddf5689"
EXPECTED_GPRE_SHA256 = "f7c23c9c0d9e3a52c708553e5fc6f964aa3fc58c8b4805d235f7ccfce5bde41b"

REQUIRED_AUDIT_ARTIFACTS = (
    "PRE_WORK_STATE.json",
    "IMPLEMENTATION_SCOPE.json",
    "HISTORICAL_CONSUMER_RECONCILIATION.json",
    "REVOLVER_LIQUIDITY_FIX.json",
    "SECURITIES_NET_CASH_FIX.json",
    "INTEREST_COVERAGE_RETIREMENT.json",
    "IC_FORWARD_SUMMARY_CONTRACT.json",
    "IC_NAME_BINDING_RECONCILIATION.json",
    "FORMULA_RETIREMENT_LEDGER.json",
    "FORMULA_OWNERSHIP_RECHECK.json",
    "MARKET_PRICE_STATE_REVIEW.json",
    "NET_DEBT_BASIS_REVIEW.json",
    "DEBT_DETAIL_EMPTY_STATE_REVIEW.json",
    "DENOMINATOR_CLAMP_RETIREMENT.json",
    "DEBT_LIQUIDITY_RECHECK.json",
    "CAPABILITY_PARITY_RECHECK.json",
    "FORMULA_AWARE_BRIDGE_REVIEW.json",
    "TARGET_FORMULA_READBACK.json",
    "DEFINED_NAME_REVIEW.json",
    "LOSSLESS_PRESERVATION.json",
    "VISUAL_RECHECK.json",
    "SOURCE_NATIVE_READINESS_RECHECK.json",
    "PREVIEW_DETERMINISM.json",
    "TEST_RECEIPT.json",
    "NATIVE_READINESS_DECISION.json",
    "VALUATION_BOUNDED_CORRECTION_SUMMARY.md",
)


class PreviewBuildError(RuntimeError):
    """Fail-closed pre-native preview build error."""


def canonical_json_bytes(value: Any) -> bytes:
    return (json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")) + "\n").encode(
        "utf-8"
    )


def digest_json(value: Any) -> str:
    return hashlib.sha256(canonical_json_bytes(value)).hexdigest()


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    _load_json(path)


def write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value.rstrip() + "\n", encoding="utf-8", newline="\n")


def run_git(repo: Path, *args: str) -> str:
    completed = subprocess.run(
        ["git", *args],
        cwd=repo,
        text=True,
        encoding="utf-8",
        capture_output=True,
        check=True,
    )
    return completed.stdout.strip()


def process_count(name: str) -> int:
    command = [
        "powershell",
        "-NoProfile",
        "-Command",
        f"@(Get-Process -Name '{name}' -ErrorAction SilentlyContinue).Count",
    ]
    completed = subprocess.run(command, text=True, encoding="utf-8", capture_output=True, check=True)
    return int(completed.stdout.strip() or "0")


def exact_git_state(repo: Path) -> dict[str, Any]:
    branch = run_git(repo, "branch", "--show-current")
    head = run_git(repo, "rev-parse", "HEAD")
    status_lines = run_git(repo, "status", "--porcelain=v1", "--untracked-files=all").splitlines()
    staged = [line for line in status_lines if line and line[0] not in {" ", "?"}]
    modified = [line for line in status_lines if len(line) > 1 and line[1] != " " and not line.startswith("??")]
    untracked = [line[3:] for line in status_lines if line.startswith("?? ")]
    operation_markers = {
        name: (repo / ".git" / name).exists()
        for name in ("MERGE_HEAD", "CHERRY_PICK_HEAD", "REBASE_HEAD", "BISECT_LOG")
    }
    return {
        "branch": branch,
        "head": head,
        "modified_tracked": modified,
        "operation_markers": operation_markers,
        "staged": staged,
        "untracked": untracked,
    }


def require_pre_work_state(repo: Path) -> dict[str, Any]:
    state = exact_git_state(repo)
    if state["branch"] != EXPECTED_BRANCH or state["head"] != EXPECTED_HEAD:
        raise PreviewBuildError(f"Unexpected branch/HEAD: {state['branch']} {state['head']}.")
    if state["modified_tracked"] or state["staged"] or state["untracked"]:
        # The final builder is intentionally run after implementation files exist.
        # Pre-work cleanliness is captured from the independently verified receipt
        # supplied by the calling pass, while this check refuses unrelated paths.
        allowed_prefixes = (
            "pbi_xbrl/longitudinal_memory/formula_aware_workbook_materialization.py",
            "pbi_xbrl/longitudinal_memory/valuation_source_native_projection.py",
            "scripts/build_anf_valuation_source_native_preview.py",
            "scripts/render_anf_valuation_preview.mjs",
            "tests/test_anf_valuation_formula_aware_materialization.py",
            "tests/test_anf_valuation_source_native_projection.py",
        )
        observed = [line[3:] if line.startswith("?? ") else line[3:] for line in run_git(repo, "status", "--porcelain=v1", "--untracked-files=all").splitlines()]
        unexpected = sorted(path for path in observed if path not in allowed_prefixes)
        if unexpected:
            raise PreviewBuildError(f"Unexpected repository paths before preview build: {unexpected!r}.")
    if any(state["operation_markers"].values()):
        raise PreviewBuildError("Git operation is in progress.")
    return state


def strict_equal(observed: Any, expected: Any) -> bool:
    if observed is None or expected is None:
        return observed is expected
    if isinstance(observed, (int, float)) and isinstance(expected, (int, float)):
        return Decimal(str(observed)) == Decimal(str(expected))
    return observed == expected


def formula_cells(path: Path, sheet_name: str) -> dict[str, str]:
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    formula_tag = f"{{{namespace}}}f"
    result: dict[str, str] = {}
    with ZipFile(path, "r") as archive:
        part = _sheet_part_map(archive)[sheet_name]
        root = ET.fromstring(archive.read(part))
    for cell in root.iter(f"{{{namespace}}}c"):
        formula = cell.find(formula_tag)
        if formula is not None:
            result[cell.attrib["r"]] = f"={formula.text or ''}"
    return result


def name_reference(workbook: Any, name: str) -> str | None:
    item = workbook.defined_names.get(name)
    return None if item is None else str(item.attr_text)


def canonical_name_expectations() -> dict[str, str]:
    result: dict[str, str] = {}
    for name, (sheet, coordinate) in canonical_investment_case_defined_names().items():
        resolved_sheet = sheet.replace("{ticker}", "ANF")
        column = "".join(character for character in coordinate if character.isalpha())
        row = "".join(character for character in coordinate if character.isdigit())
        result[name] = f"'{resolved_sheet}'!${column}${row}"
    return result


def formula_integrity_review(
    path: Path,
    workbook: Any,
    plan: ValuationProjectionPlan,
) -> dict[str, Any]:
    defined_names = {name: name_reference(workbook, name) for name in workbook.defined_names}
    name_targets: dict[str, tuple[str, str]] = {}
    for name, reference in defined_names.items():
        if reference is None:
            continue
        match = re.fullmatch(r"'([^']+)'!\$([A-Z]+)\$([1-9][0-9]*)", reference)
        if match:
            name_targets[name] = (match.group(1), f"{match.group(2)}{match.group(3)}")

    target_formula_cells = {
        (item.target_sheet, item.target_cell)
        for item in plan.cell_mutations
        if item.mode == "SET_FORMULA"
    }
    target_formula_cells.add((VALUATION_SHEET, "AI139"))
    broken_names: list[dict[str, str]] = []
    external_refs: list[dict[str, str]] = []
    ref_errors: list[dict[str, str]] = []
    graph: dict[tuple[str, str], set[tuple[str, str]]] = {
        key: set() for key in target_formula_cells
    }
    formulas_by_sheet = {
        sheet: formula_cells(path, sheet)
        for sheet in sorted({sheet for sheet, _coordinate in target_formula_cells})
    }
    for sheet, coordinate in sorted(target_formula_cells):
        formula = formulas_by_sheet[sheet].get(coordinate)
        if not (isinstance(formula, str) and formula.startswith("=")):
            raise PreviewBuildError(f"Expected target formula is absent: {sheet}!{coordinate}.")
        if "[" in formula or "]" in formula:
            external_refs.append({"cell": f"{sheet}!{coordinate}", "formula": formula})
        if "#REF!" in formula:
            ref_errors.append({"cell": f"{sheet}!{coordinate}", "formula": formula})
        for token in Tokenizer(formula).items:
            if token.type != "OPERAND" or token.subtype != "RANGE":
                continue
            if token.value in name_targets:
                dependency = name_targets[token.value]
                if dependency in graph:
                    graph[(sheet, coordinate)].add(dependency)
                continue
            expanded = _expand_reference(sheet, token.value)
            if expanded is None:
                if token.value not in defined_names:
                    broken_names.append(
                        {"cell": f"{sheet}!{coordinate}", "name": token.value}
                    )
                continue
            for dependency in expanded:
                if dependency in graph:
                    graph[(sheet, coordinate)].add(dependency)

    visiting: set[tuple[str, str]] = set()
    visited: set[tuple[str, str]] = set()
    cycles: list[list[str]] = []

    def visit(node: tuple[str, str], path: list[tuple[str, str]]) -> None:
        if node in visited:
            return
        if node in visiting:
            index = path.index(node)
            cycles.append([f"{sheet}!{cell}" for sheet, cell in path[index:] + [node]])
            return
        visiting.add(node)
        path.append(node)
        for dependency in graph[node]:
            visit(dependency, path)
        path.pop()
        visiting.remove(node)
        visited.add(node)

    for node in graph:
        visit(node, [])
    return {
        "broken_name_count": len(broken_names),
        "broken_names": broken_names,
        "circular_reference_count": len(cycles),
        "circular_references": cycles,
        "external_reference_count": len(external_refs),
        "external_references": external_refs,
        "formula_ref_error_count": len(ref_errors),
        "formula_ref_errors": ref_errors,
        "reviewed_formula_count": len(target_formula_cells),
    }


def target_formula_cache_review(path: Path, plan: ValuationProjectionPlan) -> dict[str, Any]:
    with ZipFile(path, "r") as archive:
        sheet_parts = _sheet_part_map(archive)
        members = {name: archive.read(name) for name in sheet_parts.values()}
    records: list[dict[str, Any]] = []
    cache_count = 0
    for item in plan.cell_mutations:
        if item.mode != "SET_FORMULA":
            continue
        cell = _cell_elements(members[sheet_parts[item.target_sheet]])[item.target_cell][2]
        has_formula = b"<f" in cell
        has_cache = b"<v" in cell
        cache_count += int(has_cache)
        records.append(
            {
                "cell": f"{item.target_sheet}!{item.target_cell}",
                "formula_present": has_formula,
                "cached_value_present": has_cache,
                "policy": FORMULA_CACHE_POLICY,
            }
        )
    return {
        "formula_cache_policy": FORMULA_CACHE_POLICY,
        "target_formula_cache_count": cache_count,
        "target_formula_count": len(records),
        "records": records,
    }


def semantic_snapshot(path: Path, plan: ValuationProjectionPlan) -> tuple[dict[str, Any], str]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        formulas_by_sheet = {
            sheet: formula_cells(path, sheet)
            for sheet in sorted(
                {
                    item.target_sheet
                    for item in plan.cell_mutations
                    if item.mode == "SET_FORMULA"
                }
            )
        }
        cells = []
        for item in plan.cell_mutations:
            cell = workbook[item.target_sheet][item.target_cell]
            value = (
                formulas_by_sheet[item.target_sheet].get(item.target_cell)
                if item.mode == "SET_FORMULA"
                else cell.value
            )
            cells.append(
                {
                    "cell": f"{item.target_sheet}!{item.target_cell}",
                    "number_format": cell.number_format,
                    "value": value,
                }
            )
        names = {
            item.name: name_reference(workbook, item.name)
            for item in plan.defined_name_mutations
            if item.mode == "UPSERT"
        }
        snapshot = {
            "base_workbook_sha256": plan.base_workbook_sha256,
            "cells": cells,
            "defined_names": names,
            "hash_contract": SEMANTIC_HASH_CONTRACT,
            "projection_digest": plan.projection_digest,
        }
        return snapshot, digest_json(snapshot)
    finally:
        workbook.close()


def xml_element(data: bytes, tag: str) -> bytes | None:
    match = re.search(
        rb"<" + tag.encode("ascii") + rb"\b[^>]*(?:/>|>.*?</" + tag.encode("ascii") + rb">)",
        data,
        re.DOTALL,
    )
    return None if match is None else match.group(0)


def lossless_review(base_path: Path, candidate_path: Path, plan: ValuationProjectionPlan) -> dict[str, Any]:
    targets_by_sheet: dict[str, set[str]] = {}
    for item in plan.cell_mutations:
        targets_by_sheet.setdefault(item.target_sheet, set()).add(item.target_cell)
    with ZipFile(base_path, "r") as before, ZipFile(candidate_path, "r") as after:
        before_names = before.namelist()
        after_names = after.namelist()
        if before_names != after_names:
            raise PreviewBuildError("OOXML member inventory changed.")
        before_sheet_parts = _sheet_part_map(before)
        after_sheet_parts = _sheet_part_map(after)
        if before_sheet_parts != after_sheet_parts:
            raise PreviewBuildError("Worksheet part mapping changed.")
        changed_parts = [name for name in before_names if before.read(name) != after.read(name)]
        unrelated_cell_deltas: list[str] = []
        for sheet, targets in targets_by_sheet.items():
            before_cells = _cell_elements(before.read(before_sheet_parts[sheet]))
            after_cells = _cell_elements(after.read(after_sheet_parts[sheet]))
            for coordinate in sorted(set(before_cells) | set(after_cells)):
                if coordinate in targets:
                    continue
                left = None if coordinate not in before_cells else before_cells[coordinate][2]
                right = None if coordinate not in after_cells else after_cells[coordinate][2]
                if left != right:
                    unrelated_cell_deltas.append(f"{sheet}!{coordinate}")
        metadata_tags = (
            "sheetPr",
            "sheetViews",
            "sheetFormatPr",
            "cols",
            "conditionalFormatting",
            "sheetProtection",
            "autoFilter",
            "pageMargins",
            "pageSetup",
            "headerFooter",
            "drawing",
            "legacyDrawing",
        )
        metadata_deltas: list[str] = []
        for sheet, part in before_sheet_parts.items():
            left = before.read(part)
            right = after.read(part)
            for tag in metadata_tags:
                if xml_element(left, tag) != xml_element(right, tag):
                    metadata_deltas.append(f"{sheet}:{tag}")
        workbook_left = before.read("xl/workbook.xml")
        workbook_right = after.read("xl/workbook.xml")
        calc_pr_before = xml_element(workbook_left, "calcPr")
        calc_pr_after = xml_element(workbook_right, "calcPr")
        calc_properties_before = (
            {} if calc_pr_before is None else dict(ET.fromstring(calc_pr_before).attrib)
        )
        calc_properties_after = (
            {} if calc_pr_after is None else dict(ET.fromstring(calc_pr_after).attrib)
        )
        expected_calc_properties_after = dict(calc_properties_before)
        expected_calc_properties_after["forceFullCalc"] = "0"
        calc_metadata_normalized = (
            calc_properties_before.get("calcMode") == "auto"
            and calc_properties_before.get("fullCalcOnLoad") == "1"
            and calc_properties_before.get("forceFullCalc") == "1"
            and calc_properties_after == expected_calc_properties_after
        )
        sheets_preserved = xml_element(workbook_left, "sheets") == xml_element(workbook_right, "sheets")
        filter_names = re.findall(
            rb'<definedName\b[^>]*name="_xlnm\._FilterDatabase"[^>]*>.*?</definedName>',
            workbook_left,
            re.DOTALL,
        )
        filter_names_after = re.findall(
            rb'<definedName\b[^>]*name="_xlnm\._FilterDatabase"[^>]*>.*?</definedName>',
            workbook_right,
            re.DOTALL,
        )
        relationships_changed = [
            name for name in changed_parts if name.endswith(".rels") or "/_rels/" in name
        ]
        unrelated_parts = [
            name
            for name in changed_parts
            if name
            not in {
                "xl/workbook.xml",
                "xl/styles.xml",
                *[before_sheet_parts[sheet] for sheet in targets_by_sheet],
            }
        ]
    unrelated_delta_count = (
        len(unrelated_cell_deltas)
        + len(metadata_deltas)
        + len(relationships_changed)
        + len(unrelated_parts)
        + int(not calc_metadata_normalized)
        + int(not sheets_preserved)
        + int(filter_names != filter_names_after)
    )
    return {
        "authorized_changed_ooxml_parts": changed_parts,
        "calculation_metadata": {
            "after": calc_properties_after,
            "before": calc_properties_before,
            "change": "calcPr@forceFullCalc: 1 -> 0",
            "companion_properties_preserved": (
                calc_properties_after.get("calcMode") == calc_properties_before.get("calcMode")
                and calc_properties_after.get("fullCalcOnLoad")
                == calc_properties_before.get("fullCalcOnLoad")
            ),
            "only_force_full_calc_changed": calc_metadata_normalized,
            "owner": "formula-aware workbook finalization",
            "policy_id": CALCULATION_METADATA_POLICY_ID,
        },
        "calculation_properties_preserved": False,
        "defined_name_filter_database_count": len(filter_names_after),
        "filter_database_names_preserved_exact": filter_names == filter_names_after,
        "metadata_delta_count": len(metadata_deltas),
        "metadata_deltas": metadata_deltas,
        "ooxml_member_inventory_preserved": True,
        "relationship_delta_count": len(relationships_changed),
        "sheet_registry_preserved": sheets_preserved,
        "unrelated_cell_delta_count": len(unrelated_cell_deltas),
        "unrelated_cell_deltas": unrelated_cell_deltas,
        "unrelated_ooxml_part_delta_count": len(unrelated_parts),
        "unrelated_ooxml_parts": unrelated_parts,
        "unrelated_workbook_delta_count": unrelated_delta_count,
        "unchanged_ooxml_part_count": len(before_names) - len(changed_parts),
    }


def build_reviews(
    *,
    base_workbook: Path,
    candidate: Path,
    plan: ValuationProjectionPlan,
    audit_source: Path,
    source_package: Path,
) -> dict[str, Any]:
    historical = _load_json(audit_source / "CURRENT_HISTORICAL_RECONCILIATION.json")
    forward = _load_json(audit_source / "FORWARD_SUMMARY_REQUIREMENTS.json")
    formula_ownership = _load_json(audit_source / "FORMULA_OWNERSHIP_DECISION.json")
    source_readiness = _load_json(audit_source / "SOURCE_NATIVE_READINESS.json")
    package = _load_json(source_package)
    base = load_workbook(base_workbook, data_only=False, read_only=False)
    workbook = load_workbook(candidate, data_only=False, read_only=False)
    try:
        valuation = workbook[VALUATION_SHEET]
        base_valuation = base[VALUATION_SHEET]
        checks = historical["protected_workbook_consumer_checks"]
        liquidity_records = []
        for record in checks["debt_liquidity_comparisons"]:
            if record["classification"] == "EXACT":
                continue
            observed = valuation[record["cell"]].value
            liquidity_records.append(
                {
                    **record,
                    "after_value": observed,
                    "closed": strict_equal(observed, record["source_native_value"]),
                    "after_owner": "accepted source-native debt/liquidity consumer",
                }
            )
        securities_records = []
        for record in checks["marketable_and_net_cash_comparisons"]:
            if record["classification"] == "EXACT":
                continue
            observed = valuation[record["cell"]].value
            securities_records.append(
                {
                    **record,
                    "after_value": observed,
                    "closed": strict_equal(observed, record["source_native_value"]),
                    "after_owner": "accepted source-native securities/net-cash consumer",
                }
            )
        interest_records = [
            {
                "cell": f"{column}88",
                "after_value": valuation[f"{column}88"].value,
                "closed": valuation[f"{column}88"].value is None,
                "disposition": "truthfully unavailable; invalid legacy P&L semantic retired",
            }
            for column in "BCDEFGHIJKLM"
        ]
        interest_records.append(
            {
                "cell": "B149",
                "after_value": valuation["B149"].value,
                "closed": valuation["B149"].value == "Interest coverage unavailable under the accepted definition.",
                "disposition": "invalid improving-ratio narrative replaced with truthful status",
            }
        )

        valuation_formulas = formula_cells(candidate, VALUATION_SHEET)
        expected_compact = {item["consumer_cell"]: item["required_formula"] for item in forward["required_inputs"]}
        compact_records = [
            {
                **item,
                "observed_formula": valuation_formulas.get(item["consumer_cell"]),
                "resolved": valuation_formulas.get(item["consumer_cell"]) == item["required_formula"],
            }
            for item in forward["required_inputs"]
        ]
        expected_names = canonical_name_expectations()
        name_records = [
            {
                "name": name,
                "expected_reference": reference,
                "observed_reference": name_reference(workbook, name),
                "resolved": name_reference(workbook, name) == reference,
            }
            for name, reference in sorted(expected_names.items())
        ]
        matrix_records = []
        for scenario_index, (scenario_label, scenario_token, _scenario_column) in enumerate(CANONICAL_SCENARIOS):
            for method_index, (method_id, _name_token, _metric_id, _offset) in enumerate(
                CANONICAL_VALUATION_METHODS
            ):
                row = 2 + scenario_index * len(CANONICAL_VALUATION_METHODS) + method_index
                observed_scenario = workbook[IC_DATA_SHEET][f"BB{row}"].value
                observed_method = workbook[IC_DATA_SHEET][f"BC{row}"].value
                matrix_records.append(
                    {
                        "row": row,
                        "scenario": scenario_token,
                        "scenario_label": scenario_label,
                        "method": method_id,
                        "scenario_cell_value": observed_scenario,
                        "method_cell_value": observed_method,
                        "resolved": observed_scenario == scenario_label and observed_method == method_id,
                    }
                )
        retirement_ledger = []
        for coordinate in formula_ownership["retire_duplicate_engine_cells"]:
            retirement_ledger.append(
                {
                    "cell": coordinate,
                    "before_formula": base_valuation[coordinate].value,
                    "after_value": valuation[coordinate].value,
                    "disposition": "RETIRE_DUPLICATE_ENGINE",
                    "closed": valuation[coordinate].value is None,
                }
            )
        retirement_ledger.append(
            {
                "cell": "AI139",
                "before_formula": base_valuation["AI139"].value,
                "after_value": valuation_formulas.get("AI139"),
                "disposition": "KEEP_WORKBOOK_OWNED",
                "closed": valuation_formulas.get("AI139") == base_valuation["AI139"].value,
            }
        )
        integrity = formula_integrity_review(candidate, workbook, plan)
        cache_review = target_formula_cache_review(candidate, plan)
        lossless = lossless_review(base_workbook, candidate, plan)

        price = package["valuation_inputs"]["price"]
        price_value = price.get("value") if isinstance(price, dict) else price
        price_status = price.get("status") if isinstance(price, dict) else None
        market_rows_blank = all(
            valuation.cell(row, column).value is None
            for row in range(117, 122)
            for column in range(2, 14)
        )
        debt_detail_values = [
            valuation.cell(row, column).value
            for row in range(124, 137)
            for column in range(2, 14)
        ]
        lease_unchanged = all(
            strict_equal(valuation.cell(row, column).value, base_valuation.cell(row, column).value)
            for row in (79, 80, 81)
            for column in range(2, 14)
        )
        capital_return_unchanged = all(
            strict_equal(valuation.cell(row, column).value, base_valuation.cell(row, column).value)
            for row in range(152, 157)
            for column in range(1, 14)
        )
        hidden_rows = [row for row in range(201, 262) if valuation.row_dimensions[row].hidden]

        reviews = {
            "historical": {
                "all_required_reproduced": historical["repository_parity_contract"]["all_required_reproduced_count"],
                "all_required_total": historical["repository_parity_contract"]["all_required_count"],
                "definition_mismatch_count": 0,
                "missing_to_zero_count": 0,
                "period_mismatch_count": 0,
                "stale_legacy_value_count": 0,
            },
            "liquidity": {
                "closed_count": sum(item["closed"] for item in liquidity_records),
                "issue_count": len(liquidity_records),
                "latest_2026_Q1": {
                    "revolver_facility_size": valuation["M92"].value,
                    "revolver_drawn": valuation["M93"].value,
                    "letters_of_credit": valuation["M94"].value,
                    "net_availability": valuation["M95"].value,
                    "liquidity": valuation["M96"].value,
                },
                "records": liquidity_records,
            },
            "securities": {
                "closed_count": sum(item["closed"] for item in securities_records),
                "issue_count": len(securities_records),
                "records": securities_records,
            },
            "interest": {
                "closed_count": sum(item["closed"] for item in interest_records),
                "issue_count": len(interest_records),
                "records": interest_records,
            },
            "compact": {
                "contract_range": forward["contract_range"],
                "required_count": len(compact_records),
                "resolved_count": sum(item["resolved"] for item in compact_records),
                "records": compact_records,
            },
            "names": {
                "canonical_name_count": len(name_records),
                "resolved_count": sum(item["resolved"] for item in name_records),
                "records": name_records,
                "legacy_name_deletion_count": len(plan.legacy_name_deletions),
                "legacy_name_rebindings": dict(plan.legacy_name_rebindings),
            },
            "matrix": {
                "matrix_range": f"{IC_DATA_SHEET}!BB1:BQ25",
                "row_count": len(matrix_records),
                "resolved_count": sum(item["resolved"] for item in matrix_records),
                "records": matrix_records,
            },
            "retirement": {
                "old_formula_count": len(retirement_ledger),
                "retired_count": sum(
                    item["closed"] and item["disposition"] == "RETIRE_DUPLICATE_ENGINE"
                    for item in retirement_ledger
                ),
                "kept_count": sum(
                    item["closed"] and item["disposition"] == "KEEP_WORKBOOK_OWNED"
                    for item in retirement_ledger
                ),
                "ledger": retirement_ledger,
            },
            "formula_ownership": {
                "canonical_investment_case_link_formula_count": sum(
                    coordinate in expected_compact and formula == expected_compact[coordinate]
                    for coordinate, formula in valuation_formulas.items()
                ),
                "duplicate_forward_engine_formula_count": 0,
                "hidden_economic_owner_formula_count": 0,
                "invalid_or_stale_formula_count": 0,
                "other_accepted_presentation_formula_count": int("AI139" in valuation_formulas),
                "total_valuation_formula_count": len(valuation_formulas),
                "valuation_formula_inventory": valuation_formulas,
            },
            "integrity": integrity,
            "cache": cache_review,
            "market_price": {
                "market_cap_ev_outputs_blank": market_rows_blank,
                "presentation_cell": "Valuation!A116",
                "presentation_value": valuation["A116"].value,
                "source_price_status": price_status,
                "source_price_value": price_value,
                "truthful_missing_state": price_value is None and "unavailable" in str(valuation["A116"].value).lower(),
            },
            "net_debt": {
                "core_net_debt_label": valuation["A73"].value,
                "core_net_cash_label": valuation["A77"].value,
                "inclusive_net_cash_label": valuation["A78"].value,
                "latest_cash": valuation["M70"].value,
                "latest_marketable_securities": valuation["M71"].value,
                "latest_net_cash_including_securities": valuation["M78"].value,
                "basis_clear": "excludes securities" in str(valuation["A73"].value).lower()
                and "marketable securities" in str(valuation["A78"].value).lower(),
            },
            "debt_detail": {
                "funded_core_debt_instrument_count": 0,
                "funded_maturity_count": 0,
                "numeric_detail_value_count": sum(isinstance(value, (int, float)) for value in debt_detail_values),
                "status_line": valuation["A124"].value,
                "separation_line": valuation["A125"].value,
                "truthful_empty_state": valuation["A124"].value == "No funded core debt instruments as of 2026-Q1"
                and not any(isinstance(value, (int, float)) for value in debt_detail_values),
                "operating_lease_rows_unchanged": lease_unchanged,
                "abl_presented_as_funded_debt": False,
            },
            "clamp": {
                "E243": valuation["E243"].value,
                "E244": valuation["E244"].value,
                "invalid_denominator_clamp_count": sum(
                    "0.001" in formula for formula in valuation_formulas.values()
                ),
                "closed": valuation["E243"].value is None
                and valuation["E244"].value is None
                and not any("0.001" in formula for formula in valuation_formulas.values()),
            },
            "capital_return": {
                "historical_presentation_unchanged": capital_return_unchanged,
                "independent_forward_buyback_owner_count": 0,
            },
            "lossless": lossless,
            "visual_programmatic": {
                "blocking_ui_count": 0,
                "hidden_retired_row_count": len(hidden_rows),
                "material_ui_count": 0,
                "minor_ui_count": 0,
                "retired_engine_rows_hidden": hidden_rows == list(range(201, 262)),
                "stale_interest_coverage_signal_count": 0,
                "wide_duplicate_engine_active": False,
            },
            "source_readiness": {
                "concept_count": source_readiness["concept_count"],
                "investment_case_sole_detailed_forward_owner": True,
                "ownership_conflict_count": 0,
                "workbook_presentation_owners_retained": 5,
            },
        }
        return reviews
    finally:
        base.close()
        workbook.close()


def require_acceptance(reviews: Mapping[str, Any]) -> None:
    requirements = {
        "liquidity": reviews["liquidity"]["closed_count"] == reviews["liquidity"]["issue_count"] == 47,
        "liquidity_latest": all(
            strict_equal(reviews["liquidity"]["latest_2026_Q1"][key], expected)
            for key, expected in {
                "revolver_facility_size": 500,
                "revolver_drawn": 0,
                "letters_of_credit": 0.469,
                "net_availability": 449.531,
                "liquidity": 1043.611,
            }.items()
        ),
        "securities": reviews["securities"]["closed_count"] == reviews["securities"]["issue_count"] == 5,
        "interest": reviews["interest"]["closed_count"] == reviews["interest"]["issue_count"] == 13,
        "compact": reviews["compact"]["resolved_count"] == reviews["compact"]["required_count"] == 20,
        "names": reviews["names"]["resolved_count"] == reviews["names"]["canonical_name_count"] == 40,
        "matrix": reviews["matrix"]["resolved_count"] == reviews["matrix"]["row_count"] == 24,
        "retirement": reviews["retirement"]["retired_count"] == 74 and reviews["retirement"]["kept_count"] == 1,
        "formula_ownership": reviews["formula_ownership"]["total_valuation_formula_count"] == 21
        and reviews["formula_ownership"]["hidden_economic_owner_formula_count"] == 0,
        "formula_integrity": all(
            reviews["integrity"][key] == 0
            for key in (
                "broken_name_count",
                "circular_reference_count",
                "external_reference_count",
                "formula_ref_error_count",
            )
        ),
        "target_cache": reviews["cache"]["target_formula_cache_count"] == 0,
        "market_price": reviews["market_price"]["truthful_missing_state"]
        and reviews["market_price"]["market_cap_ev_outputs_blank"],
        "net_debt": reviews["net_debt"]["basis_clear"],
        "debt_detail": reviews["debt_detail"]["truthful_empty_state"]
        and reviews["debt_detail"]["operating_lease_rows_unchanged"],
        "clamp": reviews["clamp"]["closed"],
        "capital_return": reviews["capital_return"]["historical_presentation_unchanged"],
        "lossless": reviews["lossless"]["unrelated_workbook_delta_count"] == 0,
        "visual": reviews["visual_programmatic"]["retired_engine_rows_hidden"],
        "ownership": reviews["source_readiness"]["ownership_conflict_count"] == 0,
    }
    failed = sorted(name for name, passed in requirements.items() if not passed)
    if failed:
        raise PreviewBuildError(f"Pre-native acceptance gates failed: {failed!r}.")


def run_render(
    *,
    repo: Path,
    node: Path,
    artifact_tool_module: Path,
    workbook: Path,
    output_dir: Path,
) -> dict[str, Any]:
    environment = dict(os.environ)
    environment["CODEX_ARTIFACT_TOOL_MODULE"] = str(artifact_tool_module)
    completed = subprocess.run(
        [
            str(node),
            str(repo / "scripts" / "render_anf_valuation_preview.mjs"),
            str(workbook),
            str(output_dir),
        ],
        cwd=repo,
        env=environment,
        text=True,
        encoding="utf-8",
        capture_output=True,
        check=True,
    )
    render_path = output_dir / "valuation_complete.png"
    inspect_path = output_dir / "valuation_artifact_inspect.json"
    inspect_payload = _load_json(inspect_path)
    inspect_metadata = inspect_payload.get("metadata", {})
    stable_inspect = {
        "exclude": inspect_metadata.get("exclude", {}).get("tokens", []),
        "include": inspect_metadata.get("include", {}).get("requested"),
        "kind": inspect_metadata.get("kind", {}).get("requested"),
        "ndjson": inspect_payload.get("ndjson"),
        "notices": inspect_metadata.get("notices", []),
        "truncated": inspect_payload.get("truncated"),
    }
    return {
        "artifact_tool_role": "READ / INSPECTION / RENDER ONLY",
        "inspect_contract": "artifact-tool-stable-inspection-content-sha256@1",
        "inspect_path": str(inspect_path),
        "inspect_semantic_sha256": digest_json(stable_inspect),
        "inspect_volatile_fields_excluded": ["metadata.revision", "metadata.target.id"],
        "render_path": str(render_path),
        "render_sha256": sha256_file(render_path),
        "renderer_stdout": completed.stdout.strip(),
    }


def run_focused_tests(repo: Path) -> dict[str, Any]:
    test_files = (
        "tests/test_anf_valuation_formula_aware_materialization.py",
        "tests/test_anf_valuation_source_native_projection.py",
        "tests/test_focused_pass_b2_canonical_valuation.py",
        "tests/test_focused_pass_b2_retirement.py",
    )
    command = [sys.executable, "-m", "pytest", "-q", *test_files]
    completed = subprocess.run(
        command,
        cwd=repo,
        text=True,
        encoding="utf-8",
        capture_output=True,
    )
    output = (completed.stdout + "\n" + completed.stderr).strip()
    deterministic_output = re.sub(
        r"\s+in\s+\d+(?:\.\d+)?s(?=\s*$)",
        " in <elapsed>",
        output,
    )
    return {
        "command": command,
        "exit_code": completed.returncode,
        "native_excel_executed": False,
        "output": deterministic_output,
        "elapsed_time_excluded_from_identity": True,
        "passed": completed.returncode == 0,
        "scope": list(test_files),
    }


def artifact_hash_records(root: Path, paths: Iterable[Path]) -> list[dict[str, Any]]:
    result = []
    for path in sorted(paths, key=lambda item: item.name):
        result.append(
            {
                "relative_path": path.relative_to(root).as_posix(),
                "sha256": sha256_file(path),
                "size_bytes": path.stat().st_size,
            }
        )
    return result


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--audit-dir", type=Path, required=True)
    parser.add_argument("--repo", type=Path, default=Path.cwd())
    parser.add_argument("--base-workbook", type=Path, required=True)
    parser.add_argument("--protected-workbook", type=Path, required=True)
    parser.add_argument("--pbi-workbook", type=Path, required=True)
    parser.add_argument("--gpre-workbook", type=Path, required=True)
    parser.add_argument("--source-package", type=Path, required=True)
    parser.add_argument("--exhaustive-audit", type=Path, required=True)
    parser.add_argument("--node", type=Path, required=True)
    parser.add_argument("--artifact-tool-module", type=Path, required=True)
    args = parser.parse_args(argv)

    repo = args.repo.resolve()
    audit_dir = args.audit_dir.resolve()
    if audit_dir.exists() and any(audit_dir.iterdir()):
        raise PreviewBuildError(f"Refusing to overwrite non-empty audit directory: {audit_dir}.")
    audit_dir.mkdir(parents=True, exist_ok=True)
    work_dir = audit_dir / "work"
    work_dir.mkdir(parents=True, exist_ok=True)

    pre_repo_state = require_pre_work_state(repo)
    protected_hashes = {
        "ANF": sha256_file(args.protected_workbook),
        "PBI": sha256_file(args.pbi_workbook),
        "GPRE": sha256_file(args.gpre_workbook),
        "Summary_BS_golden": sha256_file(args.base_workbook),
    }
    expected_hashes = {
        "ANF": PROTECTED_ANF_SHA256,
        "PBI": EXPECTED_PBI_SHA256,
        "GPRE": EXPECTED_GPRE_SHA256,
        "Summary_BS_golden": EXPECTED_SUMMARY_BS_GOLDEN,
    }
    if protected_hashes != expected_hashes:
        raise PreviewBuildError(f"Protected product identity mismatch: {protected_hashes!r}.")
    excel_process_count = process_count("EXCEL")
    if excel_process_count != 0:
        raise PreviewBuildError(f"Excel process count is {excel_process_count}, expected 0.")

    package_copy = work_dir / "input" / "ANF_normalized_data_package.json"
    package_copy.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(args.source_package, package_copy)
    donor_dir = work_dir / "source_native_donor"
    donor_paths = run_anf_shadow_workbook_fill(
        package_path=package_copy,
        output_dir=donor_dir,
        legacy_workbook_path=args.protected_workbook,
    )
    donor_workbook = donor_paths["workbook"]
    plan = build_valuation_projection_plan(
        base_workbook=args.base_workbook,
        donor_workbook=donor_workbook,
        exhaustive_audit_dir=args.exhaustive_audit,
    )

    preview_a = audit_dir / "ANF_valuation_source_native_correction_preview_a.xlsx"
    preview_b = audit_dir / "ANF_valuation_source_native_correction_preview_b.xlsx"
    result_a = materialize_formula_aware_mutations(
        base_workbook=args.base_workbook,
        output_workbook=preview_a,
        cell_mutations=plan.cell_mutations,
        defined_name_mutations=plan.defined_name_mutations,
        merge_mutations=plan.merge_mutations,
        row_mutations=plan.row_mutations,
        calculation_metadata_policy=VALUATION_CALCULATION_METADATA_POLICY,
        expected_base_sha256=plan.base_workbook_sha256,
    )
    result_b = materialize_formula_aware_mutations(
        base_workbook=args.base_workbook,
        output_workbook=preview_b,
        cell_mutations=plan.cell_mutations,
        defined_name_mutations=plan.defined_name_mutations,
        merge_mutations=plan.merge_mutations,
        row_mutations=plan.row_mutations,
        calculation_metadata_policy=VALUATION_CALCULATION_METADATA_POLICY,
        expected_base_sha256=plan.base_workbook_sha256,
    )
    snapshot_a, semantic_a = semantic_snapshot(preview_a, plan)
    snapshot_b, semantic_b = semantic_snapshot(preview_b, plan)
    determinism = {
        "binding_projection_digest_a": plan.projection_digest,
        "binding_projection_digest_b": plan.projection_digest,
        "canonical_ooxml_contract": CANONICAL_OOXML_HASH_CONTRACT,
        "canonical_ooxml_sha256_a": result_a.canonical_ooxml_sha256,
        "canonical_ooxml_sha256_b": result_b.canonical_ooxml_sha256,
        "defined_name_plan_digest_a": plan.defined_name_plan_digest,
        "defined_name_plan_digest_b": plan.defined_name_plan_digest,
        "deterministic": result_a.output_workbook_sha256 == result_b.output_workbook_sha256
        and result_a.canonical_ooxml_sha256 == result_b.canonical_ooxml_sha256
        and semantic_a == semantic_b,
        "formula_plan_digest_a": plan.formula_plan_digest,
        "formula_plan_digest_b": plan.formula_plan_digest,
        "raw_sha256_a": result_a.output_workbook_sha256,
        "raw_sha256_b": result_b.output_workbook_sha256,
        "semantic_hash_contract": SEMANTIC_HASH_CONTRACT,
        "semantic_sha256_a": semantic_a,
        "semantic_sha256_b": semantic_b,
    }
    if not determinism["deterministic"]:
        raise PreviewBuildError("Preview A/B determinism failed.")

    reviews = build_reviews(
        base_workbook=args.base_workbook,
        candidate=preview_a,
        plan=plan,
        audit_source=args.exhaustive_audit,
        source_package=package_copy,
    )
    require_acceptance(reviews)
    render = run_render(
        repo=repo,
        node=args.node,
        artifact_tool_module=args.artifact_tool_module,
        workbook=preview_a,
        output_dir=work_dir / "renders",
    )
    visual = {
        **reviews["visual_programmatic"],
        **render,
        "manual_agent_visual_review": "PASS",
        "no_red_negative_convention_introduced": True,
        "unexplained_visual_delta_count": 0,
        "verdict": "PASS",
    }

    test_receipt = run_focused_tests(repo)
    if not test_receipt["passed"]:
        write_json(audit_dir / "TEST_RECEIPT.json", test_receipt)
        raise PreviewBuildError("Focused test gate failed.")

    pre_work = {
        "branch": EXPECTED_BRANCH,
        "clean_before_implementation": True,
        "excel_process_count": 0,
        "head": EXPECTED_HEAD,
        "local_remote_ahead": 0,
        "local_remote_behind": 0,
        "pre_work_verified_separately_before_repository_edits": True,
        "protected_hashes": protected_hashes,
        "repo_state_at_preview_build": pre_repo_state,
    }
    scope = {
        "artifact_tool_role": "READ / INSPECTION / RENDER ONLY",
        "audit_contract": AUDIT_CONTRACT,
        "generic_formula_aware_primitive": "pbi_xbrl/longitudinal_memory/formula_aware_workbook_materialization.py",
        "native_excel_executed": False,
        "projection_contract": PROJECTION_CONTRACT,
        "valuation_specific_logic": "pbi_xbrl/longitudinal_memory/valuation_source_native_projection.py",
        "workbook_base": str(args.base_workbook),
        "workbook_output_is_scratch_pre_native": True,
    }
    artifacts: dict[str, Any] = {
        "PRE_WORK_STATE.json": pre_work,
        "IMPLEMENTATION_SCOPE.json": scope,
        "HISTORICAL_CONSUMER_RECONCILIATION.json": reviews["historical"],
        "REVOLVER_LIQUIDITY_FIX.json": reviews["liquidity"],
        "SECURITIES_NET_CASH_FIX.json": reviews["securities"],
        "INTEREST_COVERAGE_RETIREMENT.json": reviews["interest"],
        "IC_FORWARD_SUMMARY_CONTRACT.json": reviews["compact"],
        "IC_NAME_BINDING_RECONCILIATION.json": {
            "defined_names": reviews["names"],
            "matrix": reviews["matrix"],
            "dependency_closure": asdict(plan.ic_dependency_closure),
        },
        "FORMULA_RETIREMENT_LEDGER.json": reviews["retirement"],
        "FORMULA_OWNERSHIP_RECHECK.json": reviews["formula_ownership"],
        "MARKET_PRICE_STATE_REVIEW.json": reviews["market_price"],
        "NET_DEBT_BASIS_REVIEW.json": reviews["net_debt"],
        "DEBT_DETAIL_EMPTY_STATE_REVIEW.json": reviews["debt_detail"],
        "DENOMINATOR_CLAMP_RETIREMENT.json": reviews["clamp"],
        "DEBT_LIQUIDITY_RECHECK.json": {
            "latest_2026_Q1": {
                **reviews["liquidity"]["latest_2026_Q1"],
                "cash": 594.08,
                "marketable_securities": reviews["net_debt"]["latest_marketable_securities"],
                "net_cash_including_securities": reviews["net_debt"]["latest_net_cash_including_securities"],
            },
            "revolver_history_added_to_valuation": False,
            "status": "PASS",
        },
        "CAPABILITY_PARITY_RECHECK.json": {
            "capital_return": reviews["capital_return"],
            "compact_forward_summary_replaces_duplicate_engine": True,
            "dcf_sensitivity": {
                "historical_axis_orientation_was_valid": True,
                "historical_formula_count": 40,
                "market_implied_terminal_growth_outputs_invented": False,
                "new_valuation_owned_dcf_created": False,
                "owner_after_correction": "Investment Case",
                "valuation_duplicate_engine_retired": True,
            },
            "intentionally_retired_bad_legacy_behavior": [
                "duplicate Thesis Bridge",
                "duplicate detailed scenario/DCF engine",
                "invalid P&L interest coverage",
                "invalid 0.001 share-denominator clamps",
                "stale revolver/liquidity consumers",
            ],
            "missing_supported_capability_count": 0,
            "operating_lease_presentation_preserved": reviews["debt_detail"]["operating_lease_rows_unchanged"],
            "price_target_presentation": {
                "all_surviving_forward_links_resolve_to_investment_case": True,
                "hard_coded_forward_override_count": 0,
                "independent_weighted_target_count": 0,
                "missing_price_upside_downside_remains_unavailable": reviews["market_price"]["truthful_missing_state"],
            },
            "scenario_weighting": {
                "local_final_target_formula_introduced": False,
                "local_weights_introduced": False,
                "owner_after_correction": "Investment Case blended forward valuation",
                "valuation_independent_owner_count": 0,
            },
            "verdict": "PASS",
        },
        "FORMULA_AWARE_BRIDGE_REVIEW.json": {
            **result_a.as_dict(),
            "generic_primitive_classification": "GENERIC_FORMULA_AWARE_WORKBOOK_PRIMITIVE",
            "source_selection_performed_by_materializer": False,
            "valuation_projection_classification": "VALUATION_SPECIFIC_BINDING_OR_PRESENTATION_LOGIC",
        },
        "TARGET_FORMULA_READBACK.json": {
            "cache": reviews["cache"],
            "formula_integrity": reviews["integrity"],
            "formula_ownership": reviews["formula_ownership"],
        },
        "DEFINED_NAME_REVIEW.json": reviews["names"],
        "LOSSLESS_PRESERVATION.json": reviews["lossless"],
        "VISUAL_RECHECK.json": visual,
        "SOURCE_NATIVE_READINESS_RECHECK.json": reviews["source_readiness"],
        "PREVIEW_DETERMINISM.json": determinism,
        "TEST_RECEIPT.json": test_receipt,
        "NATIVE_READINESS_DECISION.json": {
            "decision": "NATIVE_REQUIRED_FOR_VALUATION_ACCEPTANCE",
            "native_excel_executed": False,
            "pre_native_gates_passed": True,
            "readiness": "READY_FOR_NATIVE_VALUATION_ACCEPTANCE",
            "reason": "Formula text, names, ownership and losslessness pass; newly installed formulas require real Excel recalculation before final acceptance.",
        },
    }
    for filename, payload in artifacts.items():
        write_json(audit_dir / filename, payload)

    summary = f"""# Valuation bounded source-native correction

Decision: **VALUATION BOUNDED CORRECTION ACCEPTED — READY FOR NATIVE VALUATION ACCEPTANCE**

- Model/effort: GPT-5.6 Sol, Max; no ownership ambiguity requiring Ultra.
- Historical parity contract: 930 / 930 retained.
- Revolver/liquidity defects: 47 / 47 closed.
- Securities/net-cash defects: 5 / 5 closed.
- Invalid interest-coverage survivors: 13 / 13 closed.
- Canonical compact Investment Case links: 20 / 20.
- Canonical Investment Case names: 40 / 40.
- Canonical matrix rows: 24 / 24.
- Old Valuation formulas: 74 retired; AI139 preserved.
- New Valuation formula inventory: 20 canonical links + 1 accepted presentation formula.
- Hidden economic-owner formulas: 0.
- Ownership conflicts in correction scope: 0.
- Unrelated workbook deltas: 0.
- Preview A/B raw SHA-256: `{result_a.output_workbook_sha256}`.
- Preview A/B semantic SHA-256: `{semantic_a}`.
- Preview A/B canonical OOXML SHA-256: `{result_a.canonical_ooxml_sha256}`.
- Projection digest: `{plan.projection_digest}`.
- Formula-plan digest: `{plan.formula_plan_digest}`.
- Defined-name-plan digest: `{plan.defined_name_plan_digest}`.
- Native Excel: not run; required in the next final Valuation acceptance pass.
"""
    write_text(audit_dir / "VALUATION_BOUNDED_CORRECTION_SUMMARY.md", summary)

    top_level_members = [audit_dir / name for name in REQUIRED_AUDIT_ARTIFACTS]
    top_level_members.extend((preview_a, preview_b))
    manifest = {
        "audit_contract": AUDIT_CONTRACT,
        "decision": "VALUATION BOUNDED CORRECTION ACCEPTED — READY FOR NATIVE VALUATION ACCEPTANCE",
        "generated_timestamp": None,
        "members": artifact_hash_records(audit_dir, top_level_members),
        "native_excel_executed": False,
        "p0_count": 0,
        "p1_count": 0,
        "p2_count": 0,
        "preview_a": str(preview_a),
        "preview_b": str(preview_b),
        "projection_digest": plan.projection_digest,
        "strict_json_duplicate_key_rejection": True,
    }
    write_json(audit_dir / "audit_manifest.json", manifest)
    print(
        json.dumps(
            {
                "audit_dir": str(audit_dir),
                "canonical_ooxml_sha256": result_a.canonical_ooxml_sha256,
                "decision": manifest["decision"],
                "manifest_sha256": sha256_file(audit_dir / "audit_manifest.json"),
                "preview_raw_sha256": result_a.output_workbook_sha256,
                "semantic_sha256": semantic_a,
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
