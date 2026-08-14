"""Audit defined-name ownership in the frozen standard workbook shell."""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import absolute_coordinate, quote_sheetname, range_boundaries


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pbi_xbrl.standard_template_audit_runner import run_audit_generator

DEFAULT_TEMPLATE = ROOT / "templates" / "standard_stock_model_template.xlsx"
DEFAULT_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
DEFAULT_BINDINGS = ROOT / "docs" / "workbook_binding_map.json"
DEFAULT_JSON = ROOT / "docs" / "standard_template_defined_name_audit.json"
DEFAULT_MD = ROOT / "docs" / "standard_template_defined_name_audit.md"
COMPANY_RE = re.compile(r"(?i)\b(?:ANF|A&F|Abercrombie|Hollister|Pitney Bowes|Green Plains|GPRE|PBI|GTX)\b")


def build_audit(*, template_path: Path, manifest_path: Path, binding_map_path: Path) -> dict[str, Any]:
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    binding_payload = json.loads(binding_map_path.read_text(encoding="utf-8"))
    active_bindings = {
        str(binding["binding_id"]): binding
        for binding in binding_payload["bindings"]
        if binding.get("writable") and binding.get("planning_state", "active") == "active"
    }
    anchors = {str(anchor["anchor_id"]): anchor for anchor in manifest.get("required_anchors") or []}
    wb = load_workbook(template_path, read_only=False, data_only=False)
    try:
        formulas = [
            str(cell.value)
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        rows: list[dict[str, Any]] = []
        for name in wb.defined_names:
            defined = wb.defined_names[name]
            attr_text = str(getattr(defined, "attr_text", "") or "")
            expected = ""
            classification = "generic_named_range"
            if name in active_bindings:
                binding = active_bindings[name]
                target = str(binding.get("planner_target") or binding["target"])
                min_col, min_row, _max_col, _max_row = range_boundaries(target)
                sheet_name = str(binding["sheet"])
                coordinate = absolute_coordinate(wb[sheet_name].cell(min_row, min_col).coordinate)
                expected = f"{quote_sheetname(sheet_name)}!{coordinate}"
                classification = "active_binding_anchor"
            elif name in anchors:
                classification = "required_shell_anchor"
            elif any(re.search(rf"(?<![A-Za-z0-9_]){re.escape(str(name))}(?![A-Za-z0-9_])", formula) for formula in formulas):
                classification = "formula_dependency"
            elif "!" not in attr_text:
                classification = "unreferenced_constant_or_alias"
            rows.append(
                {
                    "name": str(name),
                    "attr_text": attr_text,
                    "classification": classification,
                    "expected_target": expected,
                    "target_matches": not expected or expected == attr_text,
                    "company_specific": bool(COMPANY_RE.search(f"{name} {attr_text}")),
                }
            )
    finally:
        wb.close()
    rows.sort(key=lambda row: row["name"])
    return {
        "version": "1.0.0",
        "template_path": template_path.resolve().relative_to(ROOT.resolve()).as_posix(),
        "summary": {
            "retained_count": len(rows),
            "classification_counts": dict(sorted(Counter(row["classification"] for row in rows).items())),
            "target_mismatch_count": sum(1 for row in rows if not row["target_matches"]),
            "company_specific_count": sum(1 for row in rows if row["company_specific"]),
        },
        "removed_by_materializer": ["ThesisBaseAdjEBITDA_FY=815.59", "FCF_Yield=Equity_FCF_Yield (unreferenced alias)"],
        "renamed_generic_contracts": {
            "ic_lower_brand_health_rows": "ic_lower_business_health_rows",
            "ic_lower_store_productivity_rows": "ic_lower_asset_productivity_rows",
        },
        "defined_names": rows,
    }


def write_audit(*, output_json: Path, output_md: Path, **kwargs: Any) -> dict[str, Any]:
    audit = build_audit(**kwargs)
    output_json.write_text(json.dumps(audit, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    summary = audit["summary"]
    lines = [
        "# Standard Template Defined-Name Audit",
        "",
        f"- Retained names: `{summary['retained_count']}`",
        f"- Target mismatches: `{summary['target_mismatch_count']}`",
        f"- Company-specific names/constants: `{summary['company_specific_count']}`",
        "- Removed constant: `ThesisBaseAdjEBITDA_FY=815.59`",
        "- Removed stale alias: `FCF_Yield=Equity_FCF_Yield` (unreferenced by the shell).",
        "- Generic contract renames: `ic_lower_brand_health_rows` -> `ic_lower_business_health_rows`; `ic_lower_store_productivity_rows` -> `ic_lower_asset_productivity_rows`.",
        "",
        "| Name | Classification | Target | Matches contract |",
        "| --- | --- | --- | --- |",
    ]
    for row in audit["defined_names"]:
        lines.append(f"| `{row['name']}` | `{row['classification']}` | `{row['attr_text']}` | {row['target_matches']} |")
    output_md.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return audit


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--binding-map", type=Path, default=DEFAULT_BINDINGS)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--output-md", type=Path, default=DEFAULT_MD)
    args = parser.parse_args()
    is_default_run = all(
        actual.resolve() == expected.resolve()
        for actual, expected in (
            (args.template, DEFAULT_TEMPLATE),
            (args.manifest, DEFAULT_MANIFEST),
            (args.binding_map, DEFAULT_BINDINGS),
            (args.output_json, DEFAULT_JSON),
            (args.output_md, DEFAULT_MD),
        )
    )
    if is_default_run and os.environ.get("STANDARD_TEMPLATE_AUDIT_ISOLATED_RUN") != "1":
        run_audit_generator(Path(__file__), root=ROOT)
        audit = json.loads(DEFAULT_JSON.read_text(encoding="utf-8"))
    else:
        audit = write_audit(
            template_path=args.template,
            manifest_path=args.manifest,
            binding_map_path=args.binding_map,
            output_json=args.output_json,
            output_md=args.output_md,
        )
    print(json.dumps(audit["summary"], indent=2))
    return 0 if not audit["summary"]["target_mismatch_count"] and not audit["summary"]["company_specific_count"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
