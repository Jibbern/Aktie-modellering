"""Audit hidden/support sheets in the frozen standard workbook shell.

This is an audit/reporting helper only. It does not implement the value-only
new-ticker runtime and it does not build ticker workbooks.
"""
from __future__ import annotations

import argparse
import json
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = ROOT / "templates" / "standard_stock_model_template.xlsx"
DEFAULT_LAB_SOURCE = ROOT / "templates" / "lab" / "ANF_template_lab.xlsx"
DEFAULT_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
DEFAULT_AUDIT_JSON = ROOT / "docs" / "standard_template_hidden_support_audit.json"
DEFAULT_AUDIT_MD = ROOT / "docs" / "standard_template_hidden_support_audit.md"

ALLOWED_HIDDEN_SHELL_SHEETS = {
    "Hidden_Value_Flags": {
        "classification": "keep_formula_dependency",
        "reason": "Valuation!AI139 uses Hidden_Value_Flags!L2:L100 as a neutral hidden-value flag lookup helper.",
    },
    "Revolver_History": {
        "classification": "keep_neutral_helper_shell",
        "reason": "Neutral debt/liquidity support shell retained with headers only; runtime fills rows from normalized debt_liquidity.",
    },
    "Debt_Tranches_Latest": {
        "classification": "keep_neutral_helper_shell",
        "reason": "Neutral debt-tranche support shell retained with headers only; runtime fills rows from normalized debt_liquidity.",
    },
    "Debt_Profile": {
        "classification": "keep_neutral_helper_shell",
        "reason": "Neutral debt profile shell retained with headers only for valuation/liquidity workflows.",
    },
    "Guidance_Normalized": {
        "classification": "keep_neutral_helper_shell",
        "reason": "Neutral guidance support shell retained with headers only; normalized_guidance owns future values.",
    },
    "Quarter_Notes": {
        "classification": "keep_neutral_helper_shell",
        "reason": "Neutral quarter-note support shell retained with headers only; runtime fills from quarter_notes.",
    },
    "Promise_Progress": {
        "classification": "keep_neutral_helper_shell",
        "reason": "Neutral promise-progress support shell retained with headers only; runtime fills from normalized_guidance evidence.",
    },
    "History_Q": {
        "classification": "keep_neutral_helper_shell",
        "reason": "Neutral quarterly history support shell retained with headers only; runtime fills from quarterly_financials.",
    },
}

SOURCE_SPECIFIC_TERMS = (
    "ANF",
    "Abercrombie",
    "Hollister",
    "Pitney Bowes",
    "Presort",
    "SendTech",
    "Green Plains",
    "45Z",
    "RIN",
    "crush margin",
    "GTX",
)
SOURCE_FILENAME_RE = re.compile(r"\b(anf|pbi|gpre|gtx)[-_/][^\s]*\.(htm|html|pdf|xlsx|xls)\b", re.I)
SOURCE_PATH_RE = re.compile(r"StockModelData[\\/]+tickers[\\/]+(ANF|PBI|GPRE|GTX)\b", re.I)
SHEET_REF_RE = re.compile(r"'([^']+)'!|(?<![A-Za-z0-9_])([A-Za-z_][A-Za-z0-9_ ]{0,60})!")


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _sheet_refs(text: str) -> set[str]:
    refs: set[str] = set()
    for match in SHEET_REF_RE.finditer(text):
        ref = (match.group(1) or match.group(2) or "").strip()
        if ref:
            refs.add(ref)
    return refs


def _used_range(ws: Any) -> str:
    return f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"


def _source_specific_match(value: Any) -> bool:
    if value in (None, ""):
        return False
    text = str(value)
    if SOURCE_FILENAME_RE.search(text) or SOURCE_PATH_RE.search(text):
        return True
    return any(re.search(r"\b" + re.escape(term) + r"\b", text, re.I) for term in SOURCE_SPECIFIC_TERMS)


def _looks_source_raw_audit_sheet(sheet_name: str) -> bool:
    lowered = sheet_name.lower()
    markers = (
        "raw",
        "audit",
        "evidence",
        "guidance",
        "quarter_notes",
        "slides",
        "sec",
        "data_",
        "info_log",
        "ocr",
        "history",
        "adjusted_metrics",
        "nongAAP".lower(),
        "investment_case_data",
        "operating_drivers_raw",
        "promise",
    )
    return any(marker in lowered for marker in markers)


def _sheet_counts(ws: Any) -> dict[str, Any]:
    non_empty = 0
    formulas = 0
    leakage = 0
    source_raw_audit_cells = 0
    samples: list[str] = []
    source_samples: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value in (None, ""):
                continue
            non_empty += 1
            if isinstance(value, str) and value.startswith("="):
                formulas += 1
            if _source_specific_match(value):
                leakage += 1
                if len(samples) < 8:
                    samples.append(f"{cell.coordinate}={str(value)[:160]}")
            if isinstance(value, str) and (
                SOURCE_FILENAME_RE.search(value)
                or SOURCE_PATH_RE.search(value)
                or "source" in value.lower()
                or "filing" in value.lower()
                or "earnings_release" in value.lower()
                or "audit" in value.lower()
            ):
                source_raw_audit_cells += 1
                if len(source_samples) < 8:
                    source_samples.append(f"{cell.coordinate}={value[:160]}")
    return {
        "used_range": _used_range(ws),
        "non_empty_cells": non_empty,
        "formula_count": formulas,
        "table_count": len(ws.tables),
        "company_source_leakage_cells": leakage,
        "company_source_leakage_samples": samples,
        "source_raw_audit_cells": source_raw_audit_cells,
        "source_raw_audit_samples": source_samples,
    }


def _formula_references(wb: Any) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    visible_refs: dict[str, list[str]] = {sheet_name: [] for sheet_name in wb.sheetnames}
    hidden_refs: dict[str, list[str]] = {sheet_name: [] for sheet_name in wb.sheetnames}
    for ws in wb.worksheets:
        is_visible = ws.sheet_state == "visible"
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                for ref in _sheet_refs(value):
                    if ref in wb.sheetnames:
                        target = visible_refs if is_visible else hidden_refs
                        target[ref].append(f"{ws.title}!{cell.coordinate}={value[:180]}")
    return visible_refs, hidden_refs


def _defined_name_references(wb: Any) -> dict[str, list[str]]:
    refs: dict[str, list[str]] = {sheet_name: [] for sheet_name in wb.sheetnames}
    for name in wb.defined_names:
        defined_name = wb.defined_names[name]
        text = getattr(defined_name, "attr_text", "") or str(defined_name)
        for ref in _sheet_refs(text):
            if ref in wb.sheetnames:
                refs[ref].append(f"{name}: {text}")
    return refs


def _data_validation_references(wb: Any) -> dict[str, list[str]]:
    refs: dict[str, list[str]] = {sheet_name: [] for sheet_name in wb.sheetnames}
    for ws in wb.worksheets:
        for validation in getattr(ws.data_validations, "dataValidation", []) or []:
            for attr in ("formula1", "formula2"):
                value = getattr(validation, attr, None)
                if not value:
                    continue
                for ref in _sheet_refs(str(value)):
                    if ref in wb.sheetnames:
                        refs[ref].append(f"{ws.title}!{validation.sqref}:{attr}={value}")
    return refs


def _missing_referenced_sheets(wb: Any) -> dict[str, list[str]]:
    sheet_names = set(wb.sheetnames)
    missing_visible_formula: list[str] = []
    missing_hidden_formula: list[str] = []
    missing_defined_name: list[str] = []
    missing_data_validation: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                for ref in _sheet_refs(value):
                    if ref not in sheet_names:
                        target = missing_visible_formula if ws.sheet_state == "visible" else missing_hidden_formula
                        target.append(f"{ws.title}!{cell.coordinate}->{ref}")
    for name in wb.defined_names:
        defined_name = wb.defined_names[name]
        text = getattr(defined_name, "attr_text", "") or str(defined_name)
        for ref in _sheet_refs(text):
            if ref not in sheet_names:
                missing_defined_name.append(f"{name}->{ref}")
    for ws in wb.worksheets:
        for validation in getattr(ws.data_validations, "dataValidation", []) or []:
            for attr in ("formula1", "formula2"):
                value = getattr(validation, attr, None)
                if not value:
                    continue
                for ref in _sheet_refs(str(value)):
                    if ref not in sheet_names:
                        missing_data_validation.append(f"{ws.title}!{validation.sqref}:{attr}->{ref}")
    return {
        "missing_visible_formula_sheets": sorted(set(missing_visible_formula)),
        "missing_hidden_formula_sheets": sorted(set(missing_hidden_formula)),
        "missing_defined_name_sheets": sorted(set(missing_defined_name)),
        "missing_data_validation_sheets": sorted(set(missing_data_validation)),
    }


def _classify(
    *,
    sheet_name: str,
    present_in_shell: bool,
    visible_refs: list[str],
    defined_name_refs: list[str],
    data_validation_refs: list[str],
    counts: dict[str, Any],
) -> tuple[str, str]:
    if sheet_name in ALLOWED_HIDDEN_SHELL_SHEETS:
        return (
            ALLOWED_HIDDEN_SHELL_SHEETS[sheet_name]["classification"],
            ALLOWED_HIDDEN_SHELL_SHEETS[sheet_name]["reason"],
        )
    if visible_refs or defined_name_refs or data_validation_refs:
        return (
            "uncertain_manual_review",
            "Sheet has workbook dependencies and is not in the approved neutral helper allow-list.",
        )
    if _looks_source_raw_audit_sheet(sheet_name) or counts.get("source_raw_audit_cells") or counts.get("company_source_leakage_cells"):
        return (
            "delete_from_shell",
            "Source/raw/audit/runtime-output sheet from the ANF lab is not part of the frozen neutral shell.",
        )
    if present_in_shell:
        return (
            "uncertain_manual_review",
            "Hidden sheet remains in shell without an approved classification.",
        )
    return (
        "delete_from_shell",
        "Unreferenced hidden lab sheet is excluded from the frozen neutral shell.",
    )


def scan_hidden_support_package(
    *,
    template_path: Path = DEFAULT_TEMPLATE,
    lab_path: Path = DEFAULT_LAB_SOURCE,
    manifest_path: Path = DEFAULT_MANIFEST,
) -> dict[str, Any]:
    manifest = _load_json(manifest_path)
    standard_visible = set(manifest["visible_sheet_order"])
    lab_standard_visible = {name.replace("{ticker}", "ANF") for name in standard_visible}
    template_wb = load_workbook(template_path, data_only=False, read_only=False)
    lab_wb = load_workbook(lab_path, data_only=False, read_only=False) if lab_path.exists() else None
    try:
        visible_refs, hidden_formula_refs = _formula_references(template_wb)
        defined_name_refs = _defined_name_references(template_wb)
        data_validation_refs = _data_validation_references(template_wb)
        package_check = _missing_referenced_sheets(template_wb)
        template_hidden = {ws.title: ws for ws in template_wb.worksheets if ws.sheet_state != "visible"}
        lab_hidden = {
            ws.title: ws
            for ws in (lab_wb.worksheets if lab_wb is not None else [])
            if ws.title not in lab_standard_visible
        }
        candidate_names = sorted(set(template_hidden) | set(lab_hidden))

        rows: list[dict[str, Any]] = []
        pre_leakage = 0
        post_leakage = 0
        for sheet_name in candidate_names:
            ws = template_hidden.get(sheet_name)
            lab_ws = lab_hidden.get(sheet_name)
            present = ws is not None
            counts = _sheet_counts(ws) if ws is not None else _sheet_counts(lab_ws)
            lab_counts = _sheet_counts(lab_ws) if lab_ws is not None else counts
            pre_leakage += int(lab_counts["company_source_leakage_cells"])
            if present:
                post_leakage += int(counts["company_source_leakage_cells"])
            classification, reason = _classify(
                sheet_name=sheet_name,
                present_in_shell=present,
                visible_refs=visible_refs.get(sheet_name, []),
                defined_name_refs=defined_name_refs.get(sheet_name, []),
                data_validation_refs=data_validation_refs.get(sheet_name, []),
                counts=counts,
            )
            rows.append(
                {
                    "sheet_name": sheet_name,
                    "present_in_shell": present,
                    "hidden_state": ws.sheet_state if ws is not None else "deleted",
                    "used_range": counts["used_range"],
                    "non_empty_cells": counts["non_empty_cells"],
                    "formula_count": counts["formula_count"],
                    "table_count": counts["table_count"],
                    "defined_name_references": defined_name_refs.get(sheet_name, []),
                    "visible_formula_references": visible_refs.get(sheet_name, []),
                    "hidden_formula_references": hidden_formula_refs.get(sheet_name, []),
                    "data_validation_references": data_validation_refs.get(sheet_name, []),
                    "contains_company_source_text": bool(counts["company_source_leakage_cells"]),
                    "company_source_leakage_cells": counts["company_source_leakage_cells"],
                    "company_source_leakage_samples": counts["company_source_leakage_samples"],
                    "contains_source_raw_audit_data": bool(counts["source_raw_audit_cells"]) or _looks_source_raw_audit_sheet(sheet_name),
                    "source_raw_audit_cells": counts["source_raw_audit_cells"],
                    "source_raw_audit_samples": counts["source_raw_audit_samples"],
                    "classification": classification,
                    "reason": reason,
                }
            )

        retained_unclassified = [
            row["sheet_name"]
            for row in rows
            if row["present_in_shell"] and row["classification"] not in {"keep_neutral_helper_shell", "keep_formula_dependency", "keep_optional_runtime_output_shell"}
        ]

        return {
            "version": "0.1.0",
            "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat(),
            "template_path": str(template_path),
            "lab_source_path": str(lab_path),
            "allowed_hidden_shell_sheets": ALLOWED_HIDDEN_SHELL_SHEETS,
            "pre_neutralization_summary": {
                "candidate_hidden_sheet_count": len(candidate_names),
                "company_source_leakage_cells": pre_leakage,
            },
            "post_neutralization_summary": {
                "hidden_sheet_count": len(template_hidden),
                "company_source_leakage_cells": post_leakage,
                "retained_unclassified_hidden_sheets": retained_unclassified,
            },
            "package_dependency_check": package_check,
            "hidden_support_sheets": rows,
        }
    finally:
        template_wb.close()
        if lab_wb is not None:
            lab_wb.close()


def _write_markdown(payload: dict[str, Any], path: Path) -> None:
    lines = [
        "# Standard Template Hidden Support Audit",
        "",
        f"Generated at: {payload['generated_at']}",
        f"Template: `{payload['template_path']}`",
        f"ANF lab source: `{payload['lab_source_path']}`",
        "",
        "## Summary",
        "",
        f"- Candidate hidden/support sheets from lab or shell: `{payload['pre_neutralization_summary']['candidate_hidden_sheet_count']}`",
        f"- Company/source leakage cells before neutralization: `{payload['pre_neutralization_summary']['company_source_leakage_cells']}`",
        f"- Hidden sheets retained in shell: `{payload['post_neutralization_summary']['hidden_sheet_count']}`",
        f"- Company/source leakage cells after neutralization: `{payload['post_neutralization_summary']['company_source_leakage_cells']}`",
        f"- Missing visible formula sheet refs: `{len(payload['package_dependency_check']['missing_visible_formula_sheets'])}`",
        f"- Missing defined-name sheet refs: `{len(payload['package_dependency_check']['missing_defined_name_sheets'])}`",
        "",
        "## Hidden Support Sheets",
        "",
        "| Sheet | Present | Classification | Non-empty | Formulas | Tables | Leakage | Reason |",
        "|---|---:|---|---:|---:|---:|---:|---|",
    ]
    for row in payload["hidden_support_sheets"]:
        lines.append(
            "| {sheet} | {present} | {classification} | {nonempty} | {formulas} | {tables} | {leakage} | {reason} |".format(
                sheet=row["sheet_name"],
                present="yes" if row["present_in_shell"] else "no",
                classification=row["classification"],
                nonempty=row["non_empty_cells"],
                formulas=row["formula_count"],
                tables=row["table_count"],
                leakage=row["company_source_leakage_cells"],
                reason=row["reason"].replace("|", "/"),
            )
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_audit(
    *,
    template_path: Path,
    lab_path: Path,
    manifest_path: Path,
    audit_json_path: Path,
    audit_md_path: Path,
) -> dict[str, Any]:
    payload = scan_hidden_support_package(
        template_path=template_path,
        lab_path=lab_path,
        manifest_path=manifest_path,
    )
    audit_json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    _write_markdown(payload, audit_md_path)
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--lab", type=Path, default=DEFAULT_LAB_SOURCE)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--audit-json", type=Path, default=DEFAULT_AUDIT_JSON)
    parser.add_argument("--audit-md", type=Path, default=DEFAULT_AUDIT_MD)
    args = parser.parse_args()

    payload = build_audit(
        template_path=args.template.resolve(),
        lab_path=args.lab.resolve(),
        manifest_path=args.manifest.resolve(),
        audit_json_path=args.audit_json.resolve(),
        audit_md_path=args.audit_md.resolve(),
    )
    print(f"hidden support audit: {args.audit_json.resolve()}")
    print(f"hidden support audit md: {args.audit_md.resolve()}")
    print(
        "hidden leakage before/after: "
        f"{payload['pre_neutralization_summary']['company_source_leakage_cells']} / "
        f"{payload['post_neutralization_summary']['company_source_leakage_cells']}"
    )
    print(
        "missing visible formula refs: "
        f"{len(payload['package_dependency_check']['missing_visible_formula_sheets'])}"
    )
    print(
        "missing defined-name refs: "
        f"{len(payload['package_dependency_check']['missing_defined_name_sheets'])}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
