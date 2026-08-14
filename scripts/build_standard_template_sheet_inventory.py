"""Build the standard-template sheet inventory and support-sheet lifecycle docs.

This is a documentation/reporting helper only. It does not implement the
value-only new-ticker runtime and it does not build ticker workbooks.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pbi_xbrl.standard_template_audit_runner import run_audit_generator
from pbi_xbrl.workbook_modules import DEFAULT_MODULE_MANIFEST, load_workbook_module_manifest, sheet_contracts

DEFAULT_TEMPLATE = ROOT / "templates" / "standard_stock_model_template.xlsx"
DEFAULT_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
DEFAULT_BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
DEFAULT_INVENTORY_JSON = ROOT / "docs" / "standard_template_sheet_inventory.json"
DEFAULT_INVENTORY_MD = ROOT / "docs" / "standard_template_sheet_inventory.md"
DEFAULT_LIFECYCLE_JSON = ROOT / "docs" / "support_sheet_lifecycle_contract.json"
DEFAULT_LIFECYCLE_MD = ROOT / "docs" / "support_sheet_lifecycle_contract.md"

VISIBLE_QA_SURFACE_CONTRACTS = [
    {
        "sheet_name": "QA_Log",
        "owner": "frozen_shell",
        "data_owner": "value_only_runtime",
        "lifecycle": "audit_output",
        "neutral_shell_required": True,
        "headers_required": ["issue_id", "severity", "rule_id", "issue_type", "section", "root_cause", "message", "suggested_action", "occurrence_count", "visibility_disposition", "promotion_blocking", "detail_ref"],
        "allowed_writable_zones": ["A2:L5000"],
        "source_of_data": "canonical issue-ledger summaries",
        "created_when": "planned before render and filled only by the value-only runtime",
        "visibility": "visible",
        "validation_rules": ["one row per stable issue_id", "full occurrences remain JSON-authoritative", "explicit overflow only"],
    },
    {
        "sheet_name": "Needs_Review",
        "owner": "frozen_shell",
        "data_owner": "value_only_runtime",
        "lifecycle": "audit_output",
        "neutral_shell_required": True,
        "headers_required": ["issue_id", "severity", "rule_id", "section", "normalized_path", "business_row_key", "message", "suggested_action", "occurrence_count", "promotion_blocking", "detail_ref"],
        "allowed_writable_zones": ["A2:K5000"],
        "source_of_data": "canonical issues with visibility_disposition=needs_review",
        "created_when": "planned before render and filled only when unresolved actionable issues exist",
        "visibility": "visible",
        "validation_rules": ["audit-only evidence excluded", "promotion blockers retained", "explicit overflow only"],
    },
    {
        "sheet_name": "QA_Checks",
        "owner": "frozen_shell",
        "data_owner": "value_only_runtime",
        "lifecycle": "audit_output",
        "neutral_shell_required": True,
        "headers_required": ["rule_id", "status", "unique_issue_count", "occurrence_count", "blocking_count", "actionable_count", "affected_sections", "interpretation", "detail_ref"],
        "allowed_writable_zones": ["A2:I5000"],
        "source_of_data": "canonical issue-ledger rule aggregates",
        "created_when": "planned before render with one row per rule_id",
        "visibility": "visible",
        "validation_rules": ["rule-level aggregation", "blocking counts reconcile to ledger", "explicit overflow only"],
    },
]
SOURCE_TICKERS = ("PBI", "GPRE", "ANF")
PHYSICALLY_RETIRED_STANDARD_SHEETS = frozenset({"Valuation_Summary", "Valuation_Grid"})

REQUIRED_SUPPORT_SHELL_SHEETS = {
    "Hidden_Value_Flags",
    "Revolver_History",
    "Debt_Tranches_Latest",
    "Debt_Profile",
    "Guidance_Normalized",
    "Quarter_Notes",
    "Promise_Progress",
    "History_Q",
}
RUNTIME_SUPPORT_SHEETS = {
    "Debt_Maturity_Ladder",
    "Debt_Buckets",
    "Debt_Recon",
    "Debt_Tranches_Q",
    "Debt_Credit_Notes",
    "Leverage_Liquidity",
    "REPORT_IS_Q",
    "REPORT_BS_Q",
    "REPORT_CF_Q",
    "Quarter_Notes_Evidence",
    "Quarter_Narrative_Data",
    "Promise_Tracker",
    "NonGAAP_Credibility",
    "Hidden_Value_Recompute",
    "Hidden_Value_Base",
    "operating_drivers_raw",
    "Adjusted_Metrics",
    "Adjustments_Breakdown",
    "Scenario_Bridge_Tax_Treatment",
    "NonGAAP_Files",
    "Scenario_Driver_Assumptions",
    "Slides_Guidance",
    "Slides_Segments",
    "Slides_Debt_Profile",
    "NonGAAP_Bridge",
    "DATA_LineItem_Map",
    "DATA_Period_Index",
    "DATA_IS_Rules",
}
RUNTIME_AUDIT_SHEETS = {
    "Guidance_Raw",
    "Promise_Evidence",
    "Quarter_Notes_Audit",
    "DATA_Facts_Long",
    "SEC_Audit_Log",
    "Info_Log",
    "OCR_Text_Log",
    "Hidden_Value_Audit",
}
OPTIONAL_SECTOR_PACK_SHEETS = {"Economics_Overlay", "Basis_Proxy_Sandbox", "economics_market_raw"}
RESERVED_RUNTIME_SHEETS = {"History_A"}
SUPPORT_HEADERS = {
    "Hidden_Value_Flags": [
        "field",
        "display_name",
        "metric_type",
        "source_policy",
        "status",
        "reason",
        "sheet",
        "target",
        "binding_id",
        "review_status",
        "reserved",
        "has_hidden_value_issue",
    ],
    "Revolver_History": ["period", "capacity", "drawn", "availability", "covenant_status", "source_ref", "status"],
    "Debt_Tranches_Latest": ["instrument", "principal", "coupon", "maturity", "secured", "source_ref", "status"],
    "Debt_Profile": ["metric", "value", "period", "unit", "source_ref", "status"],
    "Guidance_Normalized": ["metric", "horizon", "period", "value", "unit", "status", "source_ref", "notes"],
    "Quarter_Notes": ["period", "theme", "metric", "note", "source_ref", "status"],
    "Promise_Progress": ["period", "metric", "previous_guide", "current_guide", "actual", "status", "source_ref"],
    "History_Q": ["period", "metric", "value", "unit", "source_ref", "status"],
    "History_A": ["fiscal_year", "metric", "value", "unit", "source_ref", "status"],
    "Guidance_Raw": ["source_ref", "date", "metric_text", "raw_value", "snippet", "parser_status"],
    "Promise_Evidence": ["promise_id", "metric", "period", "evidence", "source_ref", "status"],
    "Quarter_Notes_Audit": ["period", "note_id", "source_ref", "parser_status", "review_status"],
    "DATA_Facts_Long": ["ticker", "taxonomy", "fact", "period", "value", "unit", "source_ref"],
}
SHEET_REF_RE = re.compile(r"'([^']+)'!|(?<![A-Za-z0-9_])([A-Za-z_][A-Za-z0-9_ ]{0,60})!")


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _default_data_root() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = parent / "StockModelData"
        if candidate.exists():
            return candidate
    return ROOT.parent / "StockModelData"


def _source_workbook_paths(data_root: Path) -> dict[str, Path]:
    output_dir = data_root / "outputs" / "Excel stock models"
    paths: dict[str, Path] = {}
    for ticker in SOURCE_TICKERS:
        xlsx = output_dir / f"{ticker}_model.xlsx"
        xlsm = output_dir / f"{ticker}_model.xlsm"
        paths[ticker] = xlsx if xlsx.exists() or not xlsm.exists() else xlsm
    return paths


def _sheet_states(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return {ws.title: ws.sheet_state for ws in wb.worksheets}
    finally:
        wb.close()


def _sheet_refs(text: str) -> set[str]:
    refs: set[str] = set()
    for match in SHEET_REF_RE.finditer(text):
        ref = (match.group(1) or match.group(2) or "").strip()
        if ref:
            refs.add(ref)
    return refs


def _visible_formula_refs(template_path: Path) -> dict[str, list[str]]:
    wb = load_workbook(template_path, read_only=False, data_only=False)
    refs: dict[str, list[str]] = {sheet: [] for sheet in wb.sheetnames}
    try:
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if not (isinstance(value, str) and value.startswith("=")):
                        continue
                    for sheet_name in _sheet_refs(value):
                        if sheet_name in refs:
                            refs[sheet_name].append(f"{ws.title}!{cell.coordinate}")
    finally:
        wb.close()
    return refs


def _classification(
    sheet_name: str,
    standard_visible: set[str],
    module_sheets: dict[str, dict[str, Any]],
    legacy_inventory: dict[str, dict[str, Any]],
) -> tuple[str, str]:
    module_contract = module_sheets.get(sheet_name)
    if module_contract:
        role = str(module_contract["role"])
        legacy_class = str(module_contract["legacy_class"])
        if role == "visible_product":
            return "standard_visible_shell_sheet", "Visible product sheet owned by the selected frozen-shell module profile."
        if legacy_class == "B":
            return "required_support_shell_sheet", "Required neutral hidden support sheet owned by a reusable module."
        if legacy_class == "C":
            return "optional_module_shell_sheet", "Reusable optional module sheet retained as a neutral inactive header-only shell."
        return "fixture_capacity_shell_sheet", "Neutral fixture-capacity sheet reserved by an explicit module contract."
    legacy = legacy_inventory.get(sheet_name)
    if legacy:
        disposition = str(legacy["disposition"])
        if disposition == "external_detail":
            return "external_detail_sheet", str(legacy["reason"])
        if disposition == "rejected_redundant":
            return "rejected_redundant_sheet", str(legacy["reason"])
        if disposition in {"union_shell", "fixture_capacity"}:
            return "legacy_module_source_sheet", f"Legacy source equivalent of {legacy.get('union_sheet')!r}; values are not copied into the shell."
    if sheet_name in standard_visible:
        return "standard_visible_shell_sheet", "Visible UI sheet owned by the frozen shell."
    if sheet_name.endswith("_Investment_Case") or sheet_name.endswith("_Investment_Case_Data"):
        return "ticker_specific_sheet", "Ticker-specific investment-case sheet name/data projection; runtime resolves from tokenized shell or normalized package."
    if sheet_name in REQUIRED_SUPPORT_SHELL_SHEETS:
        return "required_support_shell_sheet", "Neutral support shell retained with headers only; values are filled/generated by runtime later."
    if sheet_name in RUNTIME_SUPPORT_SHEETS:
        return "runtime_generated_support_sheet", "Support projection should be created from normalized package/runtime evidence, not stored with ANF/PBI/GPRE data."
    if sheet_name in RUNTIME_AUDIT_SHEETS:
        return "runtime_generated_audit_sheet", "Audit/source rows are runtime outputs and must not be inherited from the ANF lab shell."
    if sheet_name in OPTIONAL_SECTOR_PACK_SHEETS:
        return "optional_sector_pack_sheet", "GPRE-only sector/commodity pack; excluded from the default standard shell."
    return "exclude_from_standard_shell", "Not part of the standard neutral shell unless a future explicit contract promotes it."


def _lifecycle_for(
    sheet_name: str,
    classification: str,
    module_contract: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    if classification in {"required_support_shell_sheet", "optional_module_shell_sheet", "fixture_capacity_shell_sheet"}:
        headers = [str(value) for value in (module_contract or {}).get("headers") or []]
        capacity_rows = int((module_contract or {}).get("capacity_rows") or 5000)
        capacity_columns = int((module_contract or {}).get("capacity_columns") or max(1, len(headers)))
        return {
            "sheet_name": sheet_name,
            "owner": "frozen_shell",
            "lifecycle": "static_template",
            "neutral_shell_required": True,
            "headers_required": headers,
            "allowed_writable_zones": [f"A2:{get_column_letter(capacity_columns)}{capacity_rows}"],
            "source_of_data": "normalized company data package through module-owned exact-cell bindings",
            "created_when": "materialized into the union shell as a neutral hidden header-only module sheet",
            "visibility": "hidden",
            "validation_rules": [
                "headers present",
                "no source/company text below header row",
                "no raw filing rows in frozen shell",
                "activation is profile-declared",
            ],
        }
    if classification == "external_detail_sheet":
        return {
            "sheet_name": sheet_name,
            "owner": "external_normalized_json",
            "lifecycle": "external_detail",
            "neutral_shell_required": False,
            "headers_required": [],
            "allowed_writable_zones": [],
            "source_of_data": "source evidence and normalized-package detail outside the workbook",
            "created_when": "generated outside Excel and linked through concise workbook lineage references",
            "visibility": "external_only",
            "validation_rules": [
                "not stored in frozen shell",
                "material workbook values retain detail references",
            ],
        }
    if classification == "runtime_generated_support_sheet":
        return {
            "sheet_name": sheet_name,
            "owner": "value_only_runtime",
            "lifecycle": "runtime_output",
            "neutral_shell_required": False,
            "headers_required": SUPPORT_HEADERS.get(sheet_name, ["field", "value", "period", "source_ref", "status"]),
            "allowed_writable_zones": ["runtime-defined table rows"],
            "source_of_data": "normalized package plus derived/runtime calculations",
            "created_when": "only when a promoted output workbook needs the support projection",
            "visibility": "hidden",
            "validation_rules": ["documented runtime output", "no inheritance from ANF/PBI/GPRE lab sheets"],
        }
    if classification == "runtime_generated_audit_sheet":
        return {
            "sheet_name": sheet_name,
            "owner": "value_only_runtime",
            "lifecycle": "audit_output",
            "neutral_shell_required": False,
            "headers_required": SUPPORT_HEADERS.get(sheet_name, ["source_ref", "field", "message", "status"]),
            "allowed_writable_zones": ["runtime-defined audit rows"],
            "source_of_data": "parser evidence, source coverage, validation issues, and mapping gaps",
            "created_when": "only during runtime report/workbook generation",
            "visibility": "hidden",
            "validation_rules": ["runtime-generated only", "no raw source rows in frozen shell"],
        }
    if classification == "optional_sector_pack_sheet":
        return {
            "sheet_name": sheet_name,
            "owner": "optional_sector_pack",
            "lifecycle": "optional_sector_output",
            "neutral_shell_required": False,
            "headers_required": SUPPORT_HEADERS.get(sheet_name, ["metric", "period", "value", "source_ref", "status"]),
            "allowed_writable_zones": ["sector-pack-defined rows"],
            "source_of_data": "ticker profile selected sector pack",
            "created_when": "only when an explicit sector pack is selected",
            "visibility": "hidden_or_visible_by_pack_contract",
            "validation_rules": ["must not appear in default standard shell"],
        }
    return None


def build_inventory(
    *,
    template_path: Path = DEFAULT_TEMPLATE,
    manifest_path: Path = DEFAULT_MANIFEST,
    binding_map_path: Path = DEFAULT_BINDING_MAP,
    module_manifest_path: Path = DEFAULT_MODULE_MANIFEST,
    data_root: Path | None = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    data_root = data_root or _default_data_root()
    manifest = _load_json(manifest_path)
    bindings = list((_load_json(binding_map_path).get("bindings") or []))
    module_payload = load_workbook_module_manifest(module_manifest_path)
    module_sheets = sheet_contracts(module_payload)
    legacy_inventory = {
        str(row["legacy_sheet"]): dict(row)
        for row in module_payload["legacy_sheet_inventory"]
    }
    standard_visible = set(manifest["visible_sheet_order"])
    shell_states = _sheet_states(template_path)
    source_paths = _source_workbook_paths(data_root)
    source_states = {ticker: _sheet_states(path) for ticker, path in source_paths.items()}
    visible_refs = _visible_formula_refs(template_path)
    binding_sheets = {str(binding["sheet"]) for binding in bindings}

    all_sheet_names = set(shell_states)
    all_sheet_names.update(module_sheets)
    all_sheet_names.update(legacy_inventory)
    all_sheet_names.update(OPTIONAL_SECTOR_PACK_SHEETS)
    for states in source_states.values():
        all_sheet_names.update(states)
    all_sheet_names.update(RESERVED_RUNTIME_SHEETS)
    # Physically retired sheets remain in the inventory when the versioned
    # legacy contract gives them an explicit rejected/retired disposition.
    all_sheet_names.difference_update(
        PHYSICALLY_RETIRED_STANDARD_SHEETS - set(legacy_inventory)
    )

    rows: list[dict[str, Any]] = []
    lifecycle_rows: list[dict[str, Any]] = []
    union_index = {str(name): index for index, name in enumerate(module_payload["union_sheet_order"])}
    for sheet_name in sorted(all_sheet_names, key=lambda name: (union_index.get(name, 10000), name)):
        classification, reason = _classification(sheet_name, standard_visible, module_sheets, legacy_inventory)
        dependency = bool(visible_refs.get(sheet_name)) or sheet_name in binding_sheets
        runtime_fill = classification in {
            "required_support_shell_sheet",
            "runtime_generated_support_sheet",
            "runtime_generated_audit_sheet",
            "optional_sector_pack_sheet",
            "optional_module_shell_sheet",
            "fixture_capacity_shell_sheet",
        }
        module_contract = module_sheets.get(sheet_name)
        legacy = legacy_inventory.get(sheet_name, {})
        row = {
            "sheet_name": sheet_name,
            "classification": classification,
            "present_in_standard_shell": sheet_name in shell_states,
            "present_in_PBI": sheet_name in source_states.get("PBI", {}),
            "present_in_GPRE": sheet_name in source_states.get("GPRE", {}),
            "present_in_ANF": sheet_name in source_states.get("ANF", {}),
            "standard_shell_state": shell_states.get(sheet_name, "missing"),
            "PBI_state": source_states.get("PBI", {}).get(sheet_name, "missing"),
            "GPRE_state": source_states.get("GPRE", {}).get(sheet_name, "missing"),
            "ANF_state": source_states.get("ANF", {}).get(sheet_name, "missing"),
            "reason": reason,
            "visible_formula_or_binding_dependency": dependency,
            "visible_formula_references": visible_refs.get(sheet_name, []),
            "runtime_must_create_or_fill": runtime_fill,
            "module_id": str((module_contract or {}).get("module_id") or legacy.get("module_id") or ""),
            "legacy_class": str((module_contract or {}).get("legacy_class") or legacy.get("legacy_class") or ""),
            "legacy_disposition": str(legacy.get("disposition") or ""),
        }
        rows.append(row)
        lifecycle = _lifecycle_for(sheet_name, classification, module_contract)
        if lifecycle is not None:
            lifecycle_rows.append(lifecycle)

    template_label = template_path.resolve().relative_to(ROOT.resolve()).as_posix()
    module_manifest_label = (
        module_manifest_path.resolve().relative_to(ROOT.resolve()).as_posix()
    )
    source_labels = {
        ticker: f"@data_root/{path.resolve().relative_to(data_root.resolve()).as_posix()}"
        for ticker, path in source_paths.items()
    }
    inventory = {
        "version": "0.1.0",
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat(),
        "template_path": template_label,
        "module_manifest_path": module_manifest_label,
        "module_manifest_version": str(module_payload["version"]),
        "module_profile_id": str(manifest.get("module_profile", {}).get("profile_id") or ""),
        "source_workbooks": source_labels,
        "sheets": rows,
    }
    lifecycle = {
        "version": "0.1.0",
        "generated_at": inventory["generated_at"],
        "template_path": template_label,
        "support_sheets": sorted(lifecycle_rows, key=lambda row: row["sheet_name"]),
        "visible_qa_surfaces": VISIBLE_QA_SURFACE_CONTRACTS,
    }
    return inventory, lifecycle


def _write_inventory_md(path: Path, payload: dict[str, Any]) -> None:
    lines = [
        "# Standard Template Sheet Inventory",
        "",
        "Read-only comparison of the frozen shell against saved PBI/GPRE/ANF workbooks. Raw/source/audit sheets are classified as runtime outputs unless the shell needs a neutral header-only helper.",
        "",
        "| Sheet | Class | Shell | PBI | GPRE | ANF | Runtime fill/create | Reason |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in payload["sheets"]:
        lines.append(
            f"| `{row['sheet_name']}` | `{row['classification']}` | {row['standard_shell_state']} | {row['PBI_state']} | {row['GPRE_state']} | {row['ANF_state']} | {row['runtime_must_create_or_fill']} | {row['reason']} |"
        )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def _write_lifecycle_md(path: Path, payload: dict[str, Any]) -> None:
    lines = [
        "# Support Sheet Lifecycle Contract",
        "",
        "The normalized company data package is the source of values. Support/audit sheets are either neutral frozen helpers or runtime-generated projections; the shell must not inherit ANF/PBI/GPRE source data.",
        "",
        "| Sheet | Owner | Lifecycle | Neutral shell | Visibility | Created when |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for row in payload["support_sheets"]:
        lines.append(
            f"| `{row['sheet_name']}` | `{row['owner']}` | `{row['lifecycle']}` | {row['neutral_shell_required']} | {row['visibility']} | {row['created_when']} |"
        )
    lines.extend(
        [
            "",
            "## Visible QA Surfaces",
            "",
            "Full issue occurrences remain in JSON. Visible QA sheets are bounded presentation projections only.",
            "",
            "| Sheet | Data owner | Source | Writable zone | Policy |",
            "| --- | --- | --- | --- | --- |",
        ]
    )
    for row in payload["visible_qa_surfaces"]:
        lines.append(
            f"| `{row['sheet_name']}` | `{row['data_owner']}` | {row['source_of_data']} | {', '.join(row['allowed_writable_zones'])} | {'; '.join(row['validation_rules'])} |"
        )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def write_reports(
    *,
    inventory_json: Path = DEFAULT_INVENTORY_JSON,
    inventory_md: Path = DEFAULT_INVENTORY_MD,
    lifecycle_json: Path = DEFAULT_LIFECYCLE_JSON,
    lifecycle_md: Path = DEFAULT_LIFECYCLE_MD,
    **kwargs: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    inventory, lifecycle = build_inventory(**kwargs)
    inventory_json.parent.mkdir(parents=True, exist_ok=True)
    inventory_json.write_text(json.dumps(inventory, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    lifecycle_json.write_text(json.dumps(lifecycle, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    _write_inventory_md(inventory_md, inventory)
    _write_lifecycle_md(lifecycle_md, lifecycle)
    return inventory, lifecycle


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--binding-map", type=Path, default=DEFAULT_BINDING_MAP)
    parser.add_argument("--module-manifest", type=Path, default=DEFAULT_MODULE_MANIFEST)
    parser.add_argument("--data-root", type=Path, default=None)
    parser.add_argument("--inventory-json", type=Path, default=DEFAULT_INVENTORY_JSON)
    parser.add_argument("--inventory-md", type=Path, default=DEFAULT_INVENTORY_MD)
    parser.add_argument("--lifecycle-json", type=Path, default=DEFAULT_LIFECYCLE_JSON)
    parser.add_argument("--lifecycle-md", type=Path, default=DEFAULT_LIFECYCLE_MD)
    args = parser.parse_args(argv)
    default_data_root = _default_data_root()
    selected_data_root = args.data_root.resolve() if args.data_root else default_data_root.resolve()
    is_default_run = selected_data_root == default_data_root.resolve() and all(
        actual.resolve() == expected.resolve()
        for actual, expected in (
            (args.template, DEFAULT_TEMPLATE),
            (args.manifest, DEFAULT_MANIFEST),
            (args.binding_map, DEFAULT_BINDING_MAP),
            (args.module_manifest, DEFAULT_MODULE_MANIFEST),
            (args.inventory_json, DEFAULT_INVENTORY_JSON),
            (args.inventory_md, DEFAULT_INVENTORY_MD),
            (args.lifecycle_json, DEFAULT_LIFECYCLE_JSON),
            (args.lifecycle_md, DEFAULT_LIFECYCLE_MD),
        )
    )
    if is_default_run and os.environ.get("STANDARD_TEMPLATE_AUDIT_ISOLATED_RUN") != "1":
        run_audit_generator(Path(__file__), root=ROOT, data_root=selected_data_root)
        inventory = json.loads(DEFAULT_INVENTORY_JSON.read_text(encoding="utf-8"))
        lifecycle = json.loads(DEFAULT_LIFECYCLE_JSON.read_text(encoding="utf-8"))
    else:
        inventory, lifecycle = write_reports(
            template_path=args.template,
            manifest_path=args.manifest,
            binding_map_path=args.binding_map,
            module_manifest_path=args.module_manifest,
            data_root=args.data_root,
            inventory_json=args.inventory_json,
            inventory_md=args.inventory_md,
            lifecycle_json=args.lifecycle_json,
            lifecycle_md=args.lifecycle_md,
        )
    print(
        "Wrote sheet inventory and lifecycle contract: "
        f"{len(inventory['sheets'])} inventory rows, {len(lifecycle['support_sheets'])} lifecycle rows"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
