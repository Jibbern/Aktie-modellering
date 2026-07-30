"""Audit visible/package neutrality for the frozen standard workbook shell.

This is documentation and validation support only. It does not implement the
new-ticker filler runtime and it does not build or validate ticker workbooks.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, range_boundaries


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pbi_xbrl.standard_template_audit_runner import run_audit_generator
from pbi_xbrl.workbook_modules import DEFAULT_MODULE_MANIFEST, load_workbook_module_manifest, sheet_contracts

DEFAULT_TEMPLATE = ROOT / "templates" / "standard_stock_model_template.xlsx"
DEFAULT_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
DEFAULT_OUTPUT_JSON = ROOT / "docs" / "standard_template_shell_neutrality_audit.json"
DEFAULT_OUTPUT_MD = ROOT / "docs" / "standard_template_shell_neutrality_audit.md"

COMPANY_SPECIFIC_TERMS = (
    "ANF",
    "A&F",
    "Abercrombie",
    "Hollister",
    "Pitney Bowes",
    "Presort",
    "SendTech",
    "Green Plains",
    "45Z",
    "RIN",
    "crush margin",
)
COMPANY_SPECIFIC_PATTERNS = tuple(
    re.compile(pattern, re.I)
    for pattern in (
        r"\bANF\b",
        r"A&F",
        r"\bAbercrombie\b",
        r"\bHollister\b",
        r"\bPitney Bowes\b",
        r"\bPresort\b",
        r"\bSendTech\b",
        r"\bGreen Plains\b",
        r"\b45Z\b",
        r"\bRINs?\b",
        r"\bcrush margin\b",
    )
)
FIXED_DIMENSION_MEMBERS = ("Americas", "EMEA", "APAC")
FIXED_SECTOR_LABEL_PATTERNS = tuple(
    re.compile(pattern, re.I)
    for pattern in (
        r"\bclosures?\b",
        r"\bopenings?\b",
        r"\bremodels?\b",
        r"\bstores?\s*/\s*buybacks?\b",
        r"\bstores?\s*/\s*real estate\b",
        r"\bfranchise stores?\b",
        r"\bowned stores?\b",
        r"\btariffs?\b",
        r"\bERP\b",
        r"\bfreight tailwind\b",
        r"\bmarketing headwind\b",
        r"\bbrand health\b",
        r"\bbrand family\b",
        r"\bAUR\b",
        r"\bdigital\s*/\s*omnichannel\b",
        r"\bdigital sales mix\b",
        r"\bstore growth\b",
        r"\brevenue growth vs store growth\b",
        r"\bmarketing\b",
        r"\bsourcing\s*/\s*supplier\b",
        r"\bnet sales growth\b",
        r"\badjusted EPS\b",
        r"\bshare repurchases?\b",
        r"\breal estate activity\b",
    )
)
SOURCE_SPECIFIC_PATTERNS = tuple(
    re.compile(pattern, re.I)
    for pattern in (
        r"StockModelData\\tickers\\",
        r"\b10-[KQ]\b",
        r"\b8-K\b",
        r"\bearnings release\b",
        r"\btranscript\b",
        r"\bpresentation\b",
        r"\bSEC\b",
        r"\bsource-backed\b",
        r"\bsource extract\b",
        r"\bsource material\b",
        r"\bprofile fallback\b",
        r"\b20\d{2}(?:-Q[1-4]|-\d{2}-\d{2})?\b",
        r"\$\s*-?\d",
        r"\b\d+(?:\.\d+)?\s*(?:%|bps|million|billion)\b",
        r"\brevolver(?:_|\s+)capacity(?:_|\s+)change",
        r"\bbrand_family_momentum\b",
        r"\bAUR\b",
    )
)
UNIVERSAL_HEADER_TEXT = {
    "source",
    "source_ref",
    "source / note",
    "source date",
    "notes/source",
    "status",
    "severity",
    "rule_id",
    "field",
    "message",
    "suggested_action",
    "sheet",
    "section",
    "binding_id",
    "target",
}
APPROVED_GENERIC_PRODUCT_LABELS = {
    "latest-quarter adjusted eps ($/share)",
}
RED_GREEN_STATUS_TERMS = {"PASS", "WARN", "FAIL", "N/A"}
SIGNAL_FILL_COLORS = {
    "002F80ED",
    "006FA8DC",
    "009BD3F5",
    "00A63A00",
    "00D55E00",
    "00E69F00",
    "00009E73",
    "0066C2A5",
    "0056B4E9",
    "00CC79A7",
}
GRAY_BLANK_FILLS = {"00DDDDDD", "00D9D9D9", "FFD9D9D9", "FFDDDDDD"}
STATUS_OUTPUT_FILL_COLORS = {"00D9EAF7", "00F2F2F2", "00F4CCCC", "00FFF2CC"}
NEUTRAL_BLANK_FILLS = {"", "00000000", "00FFFFFF", "FFFFFFFF"}
ALLOWED_VALUATION_SIGNAL_FILL_CELLS = {
    "O7",
    "O27",
    "O48",
    "A122",
    "A137",
    "N137",
    "B192",
}
ALLOWED_VALUATION_SIGNAL_FILL_RANGES = (
    "A122:N122",
    "A152:M152",
    "B192:S192",
)
VALUATION_RUNTIME_VALUE_CONSTANT_RANGES = (
    "D194:D216",
    "E236:E240",
    "D247:D250",
    "E253:E256",
    "L248:S250",
)
NON_NEUTRAL_CLASSES = {
    "company_specific_value",
    "company_specific_text",
    "sector_specific_label",
    "dimension_member_example",
    "source_specific_text",
    "uncertain_manual_review",
}


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _coord_in_ranges(coord: str, ranges: tuple[str, ...]) -> bool:
    row_idx, col_idx = coordinate_to_tuple(coord)
    for range_ref in ranges:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        if min_row <= row_idx <= max_row and min_col <= col_idx <= max_col:
            return True
    return False


def _sheet_name(template_name: str) -> str:
    return template_name


def _is_company_specific(text: str) -> bool:
    return any(pattern.search(text) for pattern in COMPANY_SPECIFIC_PATTERNS)


def _has_fixed_dimension_member(text: str) -> bool:
    return any(re.search(r"\b" + re.escape(term) + r"\b", text, re.I) for term in FIXED_DIMENSION_MEMBERS)


def _is_sector_specific(text: str) -> bool:
    return any(pattern.search(text) for pattern in FIXED_SECTOR_LABEL_PATTERNS)


def _is_source_specific(text: str) -> bool:
    lowered = text.lower()
    if lowered in UNIVERSAL_HEADER_TEXT:
        return False
    return any(pattern.search(text) for pattern in SOURCE_SPECIFIC_PATTERNS)


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _is_placeholder(text: str) -> bool:
    stripped = text.strip()
    return stripped.startswith("[") and stripped.endswith("]")


def _classify_cell(
    sheet: str,
    coord: str,
    value: Any,
    module_headers: set[str] | None = None,
) -> tuple[str, str]:
    if _is_formula(value):
        return "formula_static", "Excel formula retained in protected/static shell structure."
    if isinstance(value, (date, datetime)):
        return "company_specific_value", "Date/as-of constants must be supplied by normalized data."
    if isinstance(value, (int, float)):
        if sheet == "Valuation":
            return "company_specific_value", "Valuation numeric constant/value must be supplied by normalized data."
        return "company_specific_value", "Numeric visible constants are treated as source-backed values in the shell."

    text = _text(value)
    lowered = text.lower()
    if not text:
        return "placeholder_slot", "Blank cell is not recorded as non-empty content."
    if _is_placeholder(text):
        return "placeholder_slot", "Generic writable/template slot."
    if lowered in UNIVERSAL_HEADER_TEXT or lowered in (module_headers or set()):
        return "universal_template_label", "Universal QA/source/status header."
    if lowered in APPROVED_GENERIC_PRODUCT_LABELS:
        return "row_label_generic", "Approved ticker-neutral investor-facing product label."
    if _is_company_specific(text):
        return "company_specific_text", "Company/source-family term must not be standard template text."
    if _is_source_specific(text):
        return "source_specific_text", "Source/evidence text belongs in normalized package, not the shell."
    if _has_fixed_dimension_member(text):
        return "dimension_member_example", "Fixed dimension member must be a generic slot."
    if _is_sector_specific(text):
        return "sector_specific_label", "Sector-specific row label must be optional sector-pack content."
    if coord[0].upper() == "A":
        return "row_label_generic", "Generic row or section label."
    if len(text) <= 80:
        return "generic_block_label", "Reusable block label or neutral UI text."
    return "universal_template_label", "Neutral long-form template text."


def _writable_ranges_by_sheet(manifest: dict[str, Any]) -> dict[str, list[dict[str, str]]]:
    return {
        str(sheet["sheet"]): list(sheet.get("writable_zones", []))
        for sheet in manifest.get("sheets", [])
    }


def _iter_writable_cells(ws: Any, manifest: dict[str, Any]) -> Any:
    for zone in _writable_ranges_by_sheet(manifest).get(ws.title, []):
        min_col, min_row, max_col, max_row = range_boundaries(str(zone["target"]))
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                yield zone, cell


def _fill_rgb(cell: Any) -> str:
    if not cell.fill or not cell.fill.fgColor or cell.fill.fgColor.type != "rgb":
        return ""
    return str(cell.fill.fgColor.rgb or "")


def scan_neutrality_workbook(
    *,
    template_path: Path = DEFAULT_TEMPLATE,
    manifest_path: Path = DEFAULT_MANIFEST,
    module_manifest_path: Path = DEFAULT_MODULE_MANIFEST,
) -> dict[str, Any]:
    manifest = _load_json(manifest_path)
    module_payload = load_workbook_module_manifest(module_manifest_path)
    module_headers = {
        str(header).strip().lower()
        for contract in sheet_contracts(module_payload).values()
        for header in contract.get("headers") or []
    }
    required_support_sheets = {
        name
        for name, contract in sheet_contracts(module_payload).items()
        if str(contract["role"]) != "visible_product"
    }
    wb = load_workbook(template_path, data_only=False, read_only=False)
    try:
        visible_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        retained_hidden_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]
        classifications: list[dict[str, Any]] = []
        non_neutral_items: list[dict[str, Any]] = []
        classification_counts: Counter[str] = Counter()
        valuation_numeric_count = 0
        fixed_dimension_count = 0
        visible_company_source_text_count = 0

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if value in (None, ""):
                        continue
                    classification, reason = _classify_cell(
                        ws.title,
                        cell.coordinate,
                        value,
                        module_headers,
                    )
                    classification_counts[classification] += 1
                    if ws.title == "Valuation" and isinstance(value, (int, float)):
                        valuation_numeric_count += 1
                    if isinstance(value, str) and _has_fixed_dimension_member(value):
                        fixed_dimension_count += 1
                    if ws.sheet_state == "visible" and classification in {"company_specific_text", "source_specific_text"}:
                        visible_company_source_text_count += 1
                    record = {
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "value": _text(value)[:240],
                        "classification": classification,
                        "reason": reason,
                        "hidden_state": ws.sheet_state,
                    }
                    classifications.append(record)
                    if classification in NON_NEUTRAL_CLASSES:
                        non_neutral_items.append(record)

        signal_items: list[dict[str, Any]] = []
        gray_fill_items: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            for zone, cell in _iter_writable_cells(ws, manifest):
                if cell.value not in (None, ""):
                    continue
                fill = _fill_rgb(cell)
                if fill in SIGNAL_FILL_COLORS:
                    signal_items.append(
                        {
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "shell_zone": zone["zone_id"],
                            "fill": fill,
                            "classification": "heatmap_signal_without_data",
                            "reason": "Blank writable slot retains data-signal heatmap fill.",
                        }
                    )
                if fill in GRAY_BLANK_FILLS:
                    gray_fill_items.append(
                        {
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "shell_zone": zone["zone_id"],
                            "fill": fill,
                            "classification": "gray_fill_without_data",
                            "reason": "Blank writable slot retains gray data/output fill.",
                        }
                    )

        valuation_signal_items: list[dict[str, Any]] = []
        if "Valuation" in wb.sheetnames:
            valuation_ws = wb["Valuation"]
            for row in valuation_ws.iter_rows(min_row=6, max_row=valuation_ws.max_row):
                for cell in row:
                    if cell.coordinate in ALLOWED_VALUATION_SIGNAL_FILL_CELLS or _coord_in_ranges(cell.coordinate, ALLOWED_VALUATION_SIGNAL_FILL_RANGES):
                        continue
                    fill = _fill_rgb(cell)
                    if fill in SIGNAL_FILL_COLORS:
                        valuation_signal_items.append(
                            {
                                "sheet": valuation_ws.title,
                                "cell": cell.coordinate,
                                "fill": fill,
                                "classification": "valuation_signal_fill",
                                "reason": "Valuation visible shell retains a heatmap/status fill without normalized input.",
                            }
                        )

        blank_status_fill_items: list[dict[str, Any]] = []
        if "Valuation" in wb.sheetnames:
            valuation_ws = wb["Valuation"]
            for range_ref in ("B170:I188", "U51:U62"):
                min_col, min_row, max_col, max_row = range_boundaries(range_ref)
                for row in valuation_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        if cell.value not in (None, ""):
                            continue
                        fill = _fill_rgb(cell)
                        if fill not in NEUTRAL_BLANK_FILLS:
                            blank_status_fill_items.append(
                                {
                                    "sheet": valuation_ws.title,
                                    "cell": cell.coordinate,
                                    "fill": fill,
                                    "classification": "blank_status_or_value_fill",
                                    "reason": "Blank valuation status/value slot retains a result-color fill without normalized input.",
                                }
                            )

        visible_gray_fill_items: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value not in (None, ""):
                        continue
                    fill = _fill_rgb(cell)
                    if fill in GRAY_BLANK_FILLS:
                        visible_gray_fill_items.append(
                            {
                                "sheet": ws.title,
                                "cell": cell.coordinate,
                                "fill": fill,
                                "classification": "visible_blank_gray_fill",
                                "reason": "Visible blank cell retains gray data/output fill.",
                            }
                        )

        red_green_items: list[dict[str, Any]] = []
        if "Valuation" in wb.sheetnames:
            valuation_ws = wb["Valuation"]
            for row in valuation_ws.iter_rows(min_row=170, max_row=188, min_col=2, max_col=min(valuation_ws.max_column, 13)):
                for cell in row:
                    text = _text(cell.value)
                    if not text:
                        continue
                    if text in RED_GREEN_STATUS_TERMS or any(term in text for term in ("CFO/NI", "FCF TTM", "Net debt YoY", "Shares YoY")):
                        red_green_items.append(
                            {
                                "sheet": valuation_ws.title,
                                "cell": cell.coordinate,
                                "value": text[:240],
                                "classification": "red_green_status_output",
                                "reason": "Frozen shell must not retain calculated PASS/WARN/FAIL or ANF-derived flag explanations.",
                            }
                        )

        visible_runtime_value_items: list[dict[str, Any]] = []
        if "Valuation" in wb.sheetnames:
            valuation_ws = wb["Valuation"]
            for range_ref in VALUATION_RUNTIME_VALUE_CONSTANT_RANGES:
                min_col, min_row, max_col, max_row = range_boundaries(range_ref)
                for row in valuation_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        value = cell.value
                        if value in (None, ""):
                            continue
                        if _is_formula(value):
                            continue
                        visible_runtime_value_items.append(
                            {
                                "sheet": valuation_ws.title,
                                "cell": cell.coordinate,
                                "value": _text(value)[:240],
                                "classification": "visible_value_date_status_constant",
                                "reason": "Runtime value/date/status/input/output cells must be blank in the frozen shell.",
                            }
                        )

        hidden_sheets = {ws.title for ws in wb.worksheets if ws.sheet_state != "visible"}
        missing_required_support = sorted(required_support_sheets - hidden_sheets)

        summary = {
            "company_specific_value_count": classification_counts["company_specific_value"],
            "company_specific_text_count": classification_counts["company_specific_text"],
            "sector_specific_label_count": classification_counts["sector_specific_label"],
            "fixed_dimension_member_count": fixed_dimension_count,
            "source_specific_text_count": classification_counts["source_specific_text"],
            "valuation_numeric_constant_count": valuation_numeric_count,
            "signal_fill_without_value_count": len(signal_items),
            "blank_writable_non_neutral_fill_count": len(gray_fill_items),
            "visible_blank_gray_fill_count": len(visible_gray_fill_items),
            "valuation_signal_fill_count": len(valuation_signal_items),
            "blank_status_or_value_fill_count": len(blank_status_fill_items),
            "red_green_status_output_count": len(red_green_items),
            "visible_value_date_status_constant_count": len(visible_runtime_value_items),
            "visible_company_source_text_count": visible_company_source_text_count,
            "missing_required_support_shell_sheet_count": len(missing_required_support),
            "uncertain_manual_review_count": classification_counts["uncertain_manual_review"],
            "non_neutral_item_count": len(non_neutral_items) + len(signal_items) + len(gray_fill_items) + len(visible_gray_fill_items) + len(valuation_signal_items) + len(blank_status_fill_items) + len(red_green_items) + len(visible_runtime_value_items) + len(missing_required_support),
        }
        return {
            "version": "0.1.0",
            "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
            "template_path": str(template_path),
            "module_manifest_path": str(module_manifest_path),
            "visible_sheets": visible_sheets,
            "retained_hidden_sheets": retained_hidden_sheets,
            "classification_counts": dict(sorted(classification_counts.items())),
            "post_neutrality_summary": summary,
            "cell_classifications": classifications,
            "non_neutral_items": non_neutral_items,
            "style_signal_items": signal_items,
            "gray_fill_items": gray_fill_items,
            "visible_gray_fill_items": visible_gray_fill_items,
            "valuation_signal_items": valuation_signal_items,
            "blank_status_fill_items": blank_status_fill_items,
            "red_green_status_items": red_green_items,
            "visible_value_date_status_constant_items": visible_runtime_value_items,
            "missing_required_support_shell_sheets": missing_required_support,
        }
    finally:
        wb.close()


def _write_markdown(path: Path, payload: dict[str, Any]) -> None:
    summary = payload["post_neutrality_summary"]
    lines = [
        "# Standard Template Shell Neutrality Audit",
        "",
        "This audit scans the frozen standard shell for visible and retained-hidden content that would make the template company-specific or sector-specific.",
        "",
        f"- Template: `{payload['template_path']}`",
        f"- Generated at: `{payload['generated_at']}`",
        f"- Visible sheets: {', '.join(payload['visible_sheets'])}",
        f"- Retained hidden sheets: {', '.join(payload['retained_hidden_sheets']) or '-'}",
        "",
        "## Post-Neutrality Summary",
        "",
        "| Metric | Count |",
        "| --- | ---: |",
    ]
    for key, value in summary.items():
        lines.append(f"| `{key}` | {value} |")

    lines.extend(["", "## Classification Counts", "", "| Classification | Count |", "| --- | ---: |"])
    for key, value in payload["classification_counts"].items():
        lines.append(f"| `{key}` | {value} |")

    lines.extend(["", "## Remaining Non-Neutral Items", ""])
    remaining_items = [
        *payload["non_neutral_items"],
        *payload["style_signal_items"],
        *payload.get("gray_fill_items", []),
        *payload.get("visible_gray_fill_items", []),
        *payload.get("valuation_signal_items", []),
        *payload.get("blank_status_fill_items", []),
        *payload.get("red_green_status_items", []),
        *payload.get("visible_value_date_status_constant_items", []),
    ]
    if remaining_items or payload.get("missing_required_support_shell_sheets"):
        for item in remaining_items[:200]:
            lines.append(
                f"- `{item['sheet']}!{item['cell']}` `{item['classification']}`: {item.get('value') or item.get('fill')} - {item['reason']}"
            )
        for sheet_name in payload.get("missing_required_support_shell_sheets", []):
            lines.append(f"- `{sheet_name}` `missing_required_support_shell_sheet`: required hidden neutral support shell is missing.")
    else:
        lines.append("No remaining non-neutral items found.")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def build_audit(
    *,
    template_path: Path = DEFAULT_TEMPLATE,
    manifest_path: Path = DEFAULT_MANIFEST,
    module_manifest_path: Path = DEFAULT_MODULE_MANIFEST,
    output_json: Path = DEFAULT_OUTPUT_JSON,
    output_md: Path = DEFAULT_OUTPUT_MD,
) -> dict[str, Any]:
    payload = scan_neutrality_workbook(
        template_path=template_path,
        manifest_path=manifest_path,
        module_manifest_path=module_manifest_path,
    )
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    _write_markdown(output_md, payload)
    return payload


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--module-manifest", type=Path, default=DEFAULT_MODULE_MANIFEST)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON)
    parser.add_argument("--output-md", type=Path, default=DEFAULT_OUTPUT_MD)
    args = parser.parse_args(argv)

    is_default_run = all(
        actual.resolve() == expected.resolve()
        for actual, expected in (
            (args.template, DEFAULT_TEMPLATE),
            (args.manifest, DEFAULT_MANIFEST),
            (args.module_manifest, DEFAULT_MODULE_MANIFEST),
            (args.output_json, DEFAULT_OUTPUT_JSON),
            (args.output_md, DEFAULT_OUTPUT_MD),
        )
    )
    if is_default_run and os.environ.get("STANDARD_TEMPLATE_AUDIT_ISOLATED_RUN") != "1":
        run_audit_generator(Path(__file__), root=ROOT)
        payload = json.loads(DEFAULT_OUTPUT_JSON.read_text(encoding="utf-8"))
    else:
        payload = build_audit(
            template_path=args.template,
            manifest_path=args.manifest,
            module_manifest_path=args.module_manifest,
            output_json=args.output_json,
            output_md=args.output_md,
        )
    summary = payload["post_neutrality_summary"]
    print(f"neutrality audit: {args.output_json}")
    print(f"neutrality audit md: {args.output_md}")
    print(f"non-neutral items: {summary['non_neutral_item_count']}")
    print(f"valuation numeric constants: {summary['valuation_numeric_constant_count']}")
    print(f"fixed sector labels: {summary['sector_specific_label_count']}")
    print(f"fixed dimension members: {summary['fixed_dimension_member_count']}")
    print(f"signal fills without value: {summary['signal_fill_without_value_count']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
