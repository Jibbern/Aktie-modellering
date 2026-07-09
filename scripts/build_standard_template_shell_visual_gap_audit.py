"""Build a visual/template parity audit for the frozen standard shell.

This is documentation/report generation only. It does not implement the
new-ticker filler runtime and it does not build any ticker workbook.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = ROOT / "templates" / "standard_stock_model_template.xlsx"
DEFAULT_LAB = ROOT / "templates" / "lab" / "ANF_template_lab.xlsx"
DEFAULT_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
DEFAULT_AUDIT_JSON = ROOT / "docs" / "standard_template_shell_visual_gap_audit.json"
DEFAULT_AUDIT_MD = ROOT / "docs" / "standard_template_shell_visual_gap_audit.md"
DEFAULT_PREVIEW_DIR = ROOT / "templates" / "lab" / "previews"

sys.path.insert(0, str(ROOT / "scripts"))
from validate_standard_template_shell import (  # noqa: E402
    EXPECTED_STATIC_LABELS_BY_SHEET,
    MIN_STATIC_TEXT_COUNTS,
    SOURCE_SPECIFIC_TERMS,
    validate_shell,
)


PREVIEW_MODE = "openpyxl_static_not_excel_com"
MAX_PREVIEW_ROWS = 42
MAX_PREVIEW_COLS = 15
WIDE_SHEETS = {"Valuation", "Quarter_Notes_UI", "Promise_Progress_UI"}


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _sheet_name(template_name: str, ticker: str = "ANF") -> str:
    return template_name.replace("{ticker}", ticker)


def _parse_range(target: str) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(target)
    return int(min_col), int(min_row), int(max_col), int(max_row)


def _overlaps_cell(cell: Any, ranges: list[tuple[int, int, int, int]]) -> bool:
    return any(left <= cell.column <= right and top <= cell.row <= bottom for left, top, right, bottom in ranges)


def _text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _sheet_has_label(ws: Any, label: str) -> bool:
    wanted = _text(label).lower()
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and not value.startswith("="):
                found = _text(value).lower()
                if found == wanted or wanted in found:
                    return True
    return False


def _metric_counts(ws: Any, writable_ranges: list[tuple[int, int, int, int]]) -> dict[str, Any]:
    non_empty = 0
    formulas = 0
    formulas_inside_writable = 0
    formulas_outside_writable = 0
    static_template_labels = 0
    row_labels = 0
    numeric_values = 0
    text_values = 0
    source_specific_text = 0
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value in (None, ""):
                continue
            non_empty += 1
            if isinstance(value, str) and value.startswith("="):
                formulas += 1
                if _overlaps_cell(cell, writable_ranges):
                    formulas_inside_writable += 1
                else:
                    formulas_outside_writable += 1
                continue
            if isinstance(value, (int, float)):
                numeric_values += 1
            elif isinstance(value, str):
                text_values += 1
                if not _overlaps_cell(cell, writable_ranges):
                    static_template_labels += 1
                if cell.column == 1 and not _overlaps_cell(cell, writable_ranges):
                    row_labels += 1
                if any(re.search(r"\b" + re.escape(term) + r"\b", value, re.I) for term in SOURCE_SPECIFIC_TERMS):
                    source_specific_text += 1
    hidden_columns = [col for col, dim in ws.column_dimensions.items() if dim.hidden]
    return {
        "non_empty_cells": non_empty,
        "static_template_label_count": static_template_labels,
        "row_label_count": row_labels,
        "formula_count": formulas,
        "formula_count_inside_writable": formulas_inside_writable,
        "formula_count_outside_writable": formulas_outside_writable,
        "merge_count": len(ws.merged_cells.ranges),
        "numeric_value_count": numeric_values,
        "text_value_count": text_values,
        "hidden_columns": hidden_columns,
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else "",
        "source_specific_text_count": source_specific_text,
    }


def _blank_writable_cells(ws: Any, writable_ranges: list[tuple[int, int, int, int]]) -> dict[str, int]:
    blank = 0
    nonblank = 0
    for left, top, right, bottom in writable_ranges:
        for row in ws.iter_rows(min_row=top, max_row=bottom, min_col=left, max_col=right):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value in (None, ""):
                    blank += 1
                else:
                    nonblank += 1
    return {"blank": blank, "nonblank": nonblank, "total": blank + nonblank}


def _dimension_similarity(shell_ws: Any, lab_ws: Any, axis: str) -> float:
    if axis == "row":
        keys = set(shell_ws.row_dimensions) | set(lab_ws.row_dimensions)
        if not keys:
            return 1.0
        matches = 0
        total = 0
        for key in keys:
            shell_value = shell_ws.row_dimensions[key].height
            lab_value = lab_ws.row_dimensions[key].height
            if shell_value is None and lab_value is None:
                matches += 1
            elif shell_value is not None and lab_value is not None and abs(float(shell_value) - float(lab_value)) <= 0.75:
                matches += 1
            total += 1
        return round(matches / max(total, 1), 3)
    keys = set(shell_ws.column_dimensions) | set(lab_ws.column_dimensions)
    if not keys:
        return 1.0
    matches = 0
    total = 0
    for key in keys:
        shell_value = shell_ws.column_dimensions[key].width
        lab_value = lab_ws.column_dimensions[key].width
        if shell_value is None and lab_value is None:
            matches += 1
        elif shell_value is not None and lab_value is not None and abs(float(shell_value) - float(lab_value)) <= 0.75:
            matches += 1
        total += 1
    return round(matches / max(total, 1), 3)


def _gap(classification: str, message: str, *, severity: str = "info") -> dict[str, str]:
    return {"classification": classification, "severity": severity, "message": message}


def _sheet_report(
    *,
    sheet_name: str,
    shell_ws: Any,
    lab_ws: Any,
    manifest_sheet: dict[str, Any],
    validation_issues: list[dict[str, Any]],
) -> dict[str, Any]:
    writable_ranges = [_parse_range(str(zone["target"])) for zone in manifest_sheet["writable_zones"]]
    shell_counts = _metric_counts(shell_ws, writable_ranges)
    lab_counts = _metric_counts(lab_ws, writable_ranges)
    blank_writable = _blank_writable_cells(shell_ws, writable_ranges)
    expected_labels = EXPECTED_STATIC_LABELS_BY_SHEET.get(sheet_name, [])
    missing_labels = [label for label in expected_labels if not _sheet_has_label(shell_ws, label)]
    gaps: list[dict[str, str]] = []
    for label in missing_labels:
        gaps.append(_gap("should_keep_static_template_label", f"Missing reusable static label: {label}", severity="P1"))
    if blank_writable["nonblank"]:
        gaps.append(
            _gap(
                "should_clear_company_specific_value",
                f"Writable zones still contain {blank_writable['nonblank']} nonblank cells.",
                severity="P1",
            )
        )
    if shell_counts["source_specific_text_count"]:
        gaps.append(
            _gap(
                "should_clear_company_specific_text",
                f"Visible shell still contains {shell_counts['source_specific_text_count']} source/company-specific text cells.",
                severity="P1",
            )
        )
    min_labels = MIN_STATIC_TEXT_COUNTS.get(sheet_name, 0)
    if shell_counts["static_template_label_count"] < min_labels:
        gaps.append(
            _gap(
                "should_keep_static_template_label",
                f"Static label count {shell_counts['static_template_label_count']} is below minimum {min_labels}.",
                severity="P1",
            )
        )
    if lab_counts["formula_count_outside_writable"] and shell_counts["formula_count_outside_writable"] < int(
        lab_counts["formula_count_outside_writable"] * 0.9
    ):
        gaps.append(
            _gap(
                "should_keep_formula",
                "Formula/helper count outside writable zones is lower than the ANF lab contract. "
                f"shell={shell_counts['formula_count_outside_writable']} "
                f"lab={lab_counts['formula_count_outside_writable']}.",
                severity="P2",
            )
        )
    if _dimension_similarity(shell_ws, lab_ws, "row") < 0.85 or _dimension_similarity(shell_ws, lab_ws, "column") < 0.85:
        gaps.append(
            _gap(
                "should_keep_style_or_layout",
                "Row-height or column-width similarity is below the rich-shell threshold.",
                severity="P2",
            )
        )
    sheet_validation_issues = [issue for issue in validation_issues if issue.get("sheet") == sheet_name]
    for issue in sheet_validation_issues:
        gaps.append(_gap("uncertain_manual_review", f"Shell validator issue: {issue['rule_id']} {issue['target']}", severity=issue["severity"]))

    return {
        "sheet": sheet_name,
        "lab_sheet": lab_ws.title,
        "preview_mode": PREVIEW_MODE,
        "used_range": {
            "source_lab": f"A1:{get_column_letter(lab_ws.max_column)}{lab_ws.max_row}",
            "standard_shell": f"A1:{get_column_letter(shell_ws.max_column)}{shell_ws.max_row}",
        },
        "source_lab": lab_counts,
        "standard_shell": shell_counts,
        "blank_writable_cells": blank_writable,
        "row_height_similarity": _dimension_similarity(shell_ws, lab_ws, "row"),
        "column_width_similarity": _dimension_similarity(shell_ws, lab_ws, "column"),
        "missing_section_labels": missing_labels,
        "missing_row_labels": [],
        "missing_formula_static_helper_areas": [
            gap["message"] for gap in gaps if gap["classification"] == "should_keep_formula"
        ],
        "company_specific_anf_values_removed": blank_writable["nonblank"] == 0 and shell_counts["source_specific_text_count"] == 0,
        "major_section_header_presence": {
            label: _sheet_has_label(shell_ws, label) for label in expected_labels
        },
        "gap_classifications": gaps,
        "visually_complete": not gaps,
    }


def _cell_color(cell: Any) -> tuple[int, int, int]:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return (255, 255, 255)
    rgb = getattr(fill.fgColor, "rgb", None)
    if isinstance(rgb, str) and len(rgb) in {6, 8}:
        value = rgb[-6:]
        try:
            return tuple(int(value[i : i + 2], 16) for i in (0, 2, 4))  # type: ignore[return-value]
        except ValueError:
            return (255, 255, 255)
    return (255, 255, 255)


def _render_sheet_static(ws: Any, title: str) -> Image.Image:
    rows = min(ws.max_row, MAX_PREVIEW_ROWS)
    cols = min(ws.max_column, MAX_PREVIEW_COLS if ws.title not in WIDE_SHEETS else 18)
    font = ImageFont.load_default()
    col_widths: list[int] = []
    for col_idx in range(1, cols + 1):
        letter = get_column_letter(col_idx)
        width = ws.column_dimensions[letter].width or 10
        col_widths.append(max(44, min(120, int(float(width) * 7))))
    row_heights: list[int] = []
    for row_idx in range(1, rows + 1):
        height = ws.row_dimensions[row_idx].height or 15
        row_heights.append(max(18, min(42, int(float(height) * 1.25))))
    image = Image.new("RGB", (sum(col_widths) + 2, sum(row_heights) + 28), (255, 255, 255))
    draw = ImageDraw.Draw(image)
    draw.rectangle((0, 0, image.width, 26), fill=(31, 41, 55))
    draw.text((8, 7), title, fill=(255, 255, 255), font=font)
    y = 28
    for row_idx in range(1, rows + 1):
        x = 1
        for col_idx in range(1, cols + 1):
            cell = ws.cell(row_idx, col_idx)
            fill = _cell_color(cell)
            w = col_widths[col_idx - 1]
            h = row_heights[row_idx - 1]
            draw.rectangle((x, y, x + w, y + h), fill=fill, outline=(218, 225, 232))
            value = cell.value
            if value not in (None, "") and not isinstance(cell, MergedCell):
                text = str(value)
                if len(text) > 22:
                    text = text[:21] + "..."
                draw.text((x + 3, y + 4), text, fill=(17, 24, 39), font=font)
            x += w
        y += row_heights[row_idx - 1]
    return image


def _combine_vertical(images: list[Image.Image], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    width = max(image.width for image in images)
    height = sum(image.height for image in images) + (len(images) - 1) * 12
    out = Image.new("RGB", (width, height), (240, 243, 247))
    y = 0
    for image in images:
        out.paste(image, (0, y))
        y += image.height + 12
    out.save(path)
    return path


def _combine_pairs(pairs: list[tuple[Image.Image, Image.Image]], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    pair_images: list[Image.Image] = []
    for left, right in pairs:
        height = max(left.height, right.height)
        row = Image.new("RGB", (left.width + right.width + 10, height), (240, 243, 247))
        row.paste(left, (0, 0))
        row.paste(right, (left.width + 10, 0))
        pair_images.append(row)
    return _combine_vertical(pair_images, path)


def _build_previews(template_wb: Any, lab_wb: Any, visible_sheets: list[str], preview_dir: Path) -> dict[str, str]:
    shell_images: list[Image.Image] = []
    lab_images: list[Image.Image] = []
    pairs: list[tuple[Image.Image, Image.Image]] = []
    for sheet_name in visible_sheets:
        lab_sheet = _sheet_name(sheet_name)
        shell_image = _render_sheet_static(template_wb[sheet_name], f"Shell - {sheet_name}")
        lab_image = _render_sheet_static(lab_wb[lab_sheet], f"ANF lab - {lab_sheet}")
        shell_images.append(shell_image)
        lab_images.append(lab_image)
        pairs.append((shell_image, lab_image))
    return {
        "preview_mode": PREVIEW_MODE,
        "standard_shell_contact_sheet": str(_combine_vertical(shell_images, preview_dir / "standard_shell_contact_sheet.png")),
        "anf_template_lab_contact_sheet": str(_combine_vertical(lab_images, preview_dir / "anf_template_lab_contact_sheet.png")),
        "shell_vs_anf_contact_sheet": str(_combine_pairs(pairs, preview_dir / "shell_vs_anf_contact_sheet.png")),
    }


def _write_markdown(payload: dict[str, Any], path: Path) -> None:
    lines = [
        "# Standard Template Shell Visual Gap Audit",
        "",
        f"Generated at: {payload['generated_at']}",
        "",
        "Preview mode: openpyxl/static only. These PNGs are contact sheets for structural review, not Excel/COM-rendered visual PASS artifacts.",
        "",
        "## Summary",
        "",
        f"- Shell: `{payload['standard_shell']}`",
        f"- ANF lab source: `{payload['anf_template_lab']}`",
        f"- Shell validator status: `{payload['shell_validation']['status']}`",
        f"- Visually complete sheets: {sum(1 for report in payload['sheet_reports'] if report['visually_complete'])}/{len(payload['sheet_reports'])}",
        "",
        "## Contact Sheets",
        "",
    ]
    for key, value in payload["previews"].items():
        if key == "preview_mode":
            continue
        lines.append(f"- {key}: `{value}`")
    lines.extend(["", "## Sheet Reports", ""])
    for report in payload["sheet_reports"]:
        shell = report["standard_shell"]
        source = report["source_lab"]
        lines.extend(
            [
                f"### {report['sheet']}",
                "",
                f"- Used range: shell `{report['used_range']['standard_shell']}` vs ANF lab `{report['used_range']['source_lab']}`",
                f"- Non-empty cells: shell `{shell['non_empty_cells']}` vs ANF lab `{source['non_empty_cells']}`",
                f"- Static/template labels: shell `{shell['static_template_label_count']}` vs ANF lab `{source['static_template_label_count']}`",
                f"- Row labels: shell `{shell['row_label_count']}` vs ANF lab `{source['row_label_count']}`",
                f"- Formulas: shell `{shell['formula_count']}` vs ANF lab `{source['formula_count']}`",
                f"- Formula/helper cells outside writable zones: shell `{shell['formula_count_outside_writable']}` vs ANF lab `{source['formula_count_outside_writable']}`",
                f"- ANF formulas cleared because they were inside writable value zones: `{source['formula_count_inside_writable']}`",
                f"- Merges: shell `{shell['merge_count']}` vs ANF lab `{source['merge_count']}`",
                f"- Hidden columns: `{', '.join(shell['hidden_columns']) if shell['hidden_columns'] else 'none'}`",
                f"- Freeze panes: `{shell['freeze_panes']}`",
                f"- Row height similarity: `{report['row_height_similarity']}`; column width similarity: `{report['column_width_similarity']}`",
                f"- Writable cells blank/nonblank: `{report['blank_writable_cells']['blank']}` / `{report['blank_writable_cells']['nonblank']}`",
                f"- Visually complete: `{report['visually_complete']}`",
            ]
        )
        if report["gap_classifications"]:
            lines.append("- Gaps:")
            for gap in report["gap_classifications"]:
                lines.append(f"  - `{gap['classification']}` {gap['severity']}: {gap['message']}")
        else:
            lines.append("- Gaps: none material after clearing company-specific value zones.")
        lines.append("")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_audit(
    *,
    template_path: Path,
    lab_path: Path,
    manifest_path: Path,
    audit_json_path: Path,
    audit_md_path: Path,
    preview_dir: Path,
) -> dict[str, Any]:
    manifest = _load_json(manifest_path)
    validation = validate_shell(template_path=template_path, manifest_path=manifest_path)
    validation_issues = list(validation.get("issues") or [])
    template_wb = load_workbook(template_path, data_only=False, read_only=False)
    lab_wb = load_workbook(lab_path, data_only=False, read_only=False)
    try:
        manifest_by_sheet = {sheet["sheet"]: sheet for sheet in manifest["sheets"]}
        sheet_reports = []
        for sheet_name in manifest["visible_sheet_order"]:
            lab_sheet = _sheet_name(sheet_name)
            sheet_reports.append(
                _sheet_report(
                    sheet_name=sheet_name,
                    shell_ws=template_wb[sheet_name],
                    lab_ws=lab_wb[lab_sheet],
                    manifest_sheet=manifest_by_sheet[sheet_name],
                    validation_issues=validation_issues,
                )
            )
        previews = _build_previews(template_wb, lab_wb, list(manifest["visible_sheet_order"]), preview_dir)
    finally:
        template_wb.close()
        lab_wb.close()

    payload = {
        "version": "0.1.0",
        "generated_at": datetime.now(UTC).isoformat(),
        "preview_mode": PREVIEW_MODE,
        "standard_shell": str(template_path),
        "anf_template_lab": str(lab_path),
        "manifest": str(manifest_path),
        "shell_validation": validation,
        "previews": previews,
        "sheet_reports": sheet_reports,
        "gap_classification_legend": [
            "should_keep_static_template_label",
            "should_keep_row_label",
            "should_keep_formula",
            "should_keep_style_or_layout",
            "should_keep_hidden_helper_structure",
            "should_clear_company_specific_value",
            "should_clear_company_specific_text",
            "should_exclude_sector_specific_block",
            "uncertain_manual_review",
        ],
    }
    audit_json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    _write_markdown(payload, audit_md_path)
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--lab", type=Path, default=DEFAULT_LAB)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--audit-json", type=Path, default=DEFAULT_AUDIT_JSON)
    parser.add_argument("--audit-md", type=Path, default=DEFAULT_AUDIT_MD)
    parser.add_argument("--preview-dir", type=Path, default=DEFAULT_PREVIEW_DIR)
    args = parser.parse_args()

    payload = build_audit(
        template_path=args.template.resolve(),
        lab_path=args.lab.resolve(),
        manifest_path=args.manifest.resolve(),
        audit_json_path=args.audit_json.resolve(),
        audit_md_path=args.audit_md.resolve(),
        preview_dir=args.preview_dir.resolve(),
    )
    print(f"visual gap audit: {args.audit_json.resolve()}")
    print(f"visual gap audit md: {args.audit_md.resolve()}")
    print(f"preview mode: {payload['preview_mode']}")
    for key, value in payload["previews"].items():
        if key != "preview_mode":
            print(f"{key}: {value}")
    return 0 if payload["shell_validation"]["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
