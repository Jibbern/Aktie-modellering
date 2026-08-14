"""Build workbook block architecture docs from existing model workbooks.

This is a read-only analysis and lab-artifact script. It does not build,
patch, validate, promote, or canonicalize any ticker workbook, and it is not
the new-ticker value-only filler runtime.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, range_boundaries

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from pbi_xbrl.json_schema_validation import load_json_strict
from pbi_xbrl.standard_template_audit_runner import run_audit_generator


def _default_data_root() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = parent / "StockModelData"
        if candidate.exists():
            return candidate
    return ROOT.parent / "StockModelData"


DEFAULT_DATA_ROOT = _default_data_root()
DEFAULT_SOURCE_DIR = DEFAULT_DATA_ROOT / "outputs" / "Excel stock models"
DEFAULT_LAB_PATH = ROOT / "templates" / "lab" / "ANF_template_lab.xlsx"
ARCHITECTURE_JSON = ROOT / "docs" / "workbook_block_architecture.json"
ARCHITECTURE_MD = ROOT / "docs" / "workbook_block_architecture.md"
COVERAGE_JSON = ROOT / "docs" / "workbook_block_coverage_matrix.json"
COVERAGE_MD = ROOT / "docs" / "workbook_block_coverage_matrix.md"

SOURCE_TICKERS = ("ANF", "PBI", "GPRE")
STANDARD_VISIBLE_SHEETS = [
    "SUMMARY",
    "Valuation",
    "BS_Segments",
    "Operating_Drivers",
    "{ticker}_Investment_Case",
    "Quarter_Notes_UI",
    "Promise_Progress_UI",
    "QA_Log",
    "Needs_Review",
    "QA_Checks",
]
SECTOR_OVERLAYS = ("Economics_Overlay", "Basis_Proxy_Sandbox")
OPTIONAL_SECTOR_PACKS = [
    {
        "pack_id": "retail_operating_pack",
        "display_name": "Retail operating pack",
        "status": "optional_not_in_standard_shell",
        "example_members": ["stores", "closures", "openings", "remodels", "tariffs", "freight", "marketing"],
        "activation_rule": "Future runtime may include only when ticker profile explicitly selects retail/store-footprint drivers.",
    },
    {
        "pack_id": "commodity_ethanol_pack",
        "display_name": "Commodity/ethanol pack",
        "status": "optional_not_in_standard_shell",
        "example_members": ["crush margin", "45Z", "RINs", "corn", "natural gas", "oil"],
        "activation_rule": "Future runtime may include only when ticker profile explicitly selects commodity/ethanol economics.",
    },
    {
        "pack_id": "shipping_mail_pack",
        "display_name": "Shipping/mail pack",
        "status": "optional_not_in_standard_shell",
        "example_members": ["Presort", "SendTech", "USPS", "GEC"],
        "activation_rule": "Future runtime may include only when ticker profile explicitly selects mail/shipping operating drivers.",
    },
    {
        "pack_id": "auto_supplier_pack",
        "display_name": "Auto supplier pack",
        "status": "optional_not_in_standard_shell",
        "example_members": ["production", "turbo penetration", "BEV", "OEM mix"],
        "activation_rule": "Future runtime may include only when ticker profile explicitly selects auto-supplier drivers.",
    },
]
COMPANY_SPECIFIC_TERMS = (
    "Abercrombie",
    "Hollister",
    "Pitney Bowes",
    "Presort",
    "SendTech",
    "Green Plains",
    "45Z",
    "RIN",
    "crush margin",
    "A&F",
)


@dataclass(frozen=True)
class Bounds:
    min_col: int
    min_row: int
    max_col: int
    max_row: int


def _load_json(path: Path) -> dict[str, Any]:
    payload = load_json_strict(path)
    if not isinstance(payload, dict):
        raise ValueError(f"JSON contract must be an object: {path}")
    return payload


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _portable_template_lab_paths(
    *, source_path: Path, source_dir: Path, lab_path: Path
) -> tuple[str, str]:
    source_relative = source_path.resolve().relative_to(source_dir.resolve()).as_posix()
    lab_relative = lab_path.resolve().relative_to(ROOT.resolve()).as_posix()
    return f"@source_dir/{source_relative}", lab_relative


def _bounds(range_ref: str) -> Bounds:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    return Bounds(int(min_col), int(min_row), int(max_col), int(max_row))


def _range_ref(bounds: Bounds) -> str:
    return f"{get_column_letter(bounds.min_col)}{bounds.min_row}:{get_column_letter(bounds.max_col)}{bounds.max_row}"


def _intersects(first: Bounds, second: Bounds) -> bool:
    return not (
        first.max_col < second.min_col
        or second.max_col < first.min_col
        or first.max_row < second.min_row
        or second.max_row < first.min_row
    )


def _contains(outer: Bounds, inner: Bounds) -> bool:
    return (
        outer.min_col <= inner.min_col
        and inner.max_col <= outer.max_col
        and outer.min_row <= inner.min_row
        and inner.max_row <= outer.max_row
    )


def _cell_in_bounds(cell_ref: str, bounds: Bounds) -> bool:
    col_text = "".join(ch for ch in cell_ref if ch.isalpha())
    row_text = "".join(ch for ch in cell_ref if ch.isdigit())
    if not col_text or not row_text:
        return False
    col = 0
    for char in col_text.upper():
        col = col * 26 + (ord(char) - ord("A") + 1)
    row = int(row_text)
    return bounds.min_col <= col <= bounds.max_col and bounds.min_row <= row <= bounds.max_row


def _outside_binding_targets(records: list[dict[str, str]], bindings: list[dict[str, Any]]) -> list[dict[str, str]]:
    binding_bounds = [_bounds(str(entry["target"])) for entry in bindings]
    return [
        record
        for record in records
        if not any(_cell_in_bounds(record["cell"], target_bounds) for target_bounds in binding_bounds)
    ]


def _sheet_name(template_name: str, ticker: str) -> str:
    return template_name.replace("{ticker}", ticker)


def _text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _is_company_specific(text: str) -> bool:
    lowered = text.lower()
    return any(term.lower() in lowered for term in COMPANY_SPECIFIC_TERMS)


def _is_source_specific(text: str) -> bool:
    lowered = text.lower()
    markers = (
        "source",
        "sec",
        "10-q",
        "10-k",
        "transcript",
        "earnings call",
        "presentation",
        "press release",
        "guidance",
        "quarter",
        "fiscal",
    )
    return any(marker in lowered for marker in markers)


def _looks_numeric(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return True
    text = str(value or "").strip().replace(",", ".").replace("%", "")
    if not text or text.upper() == "N/A":
        return False
    try:
        float(text)
    except ValueError:
        return False
    return True


def _cell_value(cell: Cell) -> str:
    value = _text(cell.value)
    if len(value) > 160:
        return value[:157] + "..."
    return value


def _cell_record(cell: Cell) -> dict[str, str]:
    return {"cell": cell.coordinate, "value": _cell_value(cell)}


def _formula_record(cell: Cell) -> dict[str, str]:
    formula = _text(cell.value)
    if len(formula) > 200:
        formula = formula[:197] + "..."
    return {"cell": cell.coordinate, "formula": formula}


def _iter_range_cells(ws: Any, bounds: Bounds) -> Iterable[Cell]:
    max_row = min(bounds.max_row, ws.max_row)
    max_col = min(bounds.max_col, ws.max_column)
    if max_row < bounds.min_row or max_col < bounds.min_col:
        return []
    return (
        cell
        for row in ws.iter_rows(
            min_row=bounds.min_row,
            max_row=max_row,
            min_col=bounds.min_col,
            max_col=max_col,
        )
        for cell in row
    )


def _collect_title_cells(ws: Any, block_bounds: Bounds, anchor_labels: list[str]) -> list[dict[str, str]]:
    scan_bounds = Bounds(
        max(1, block_bounds.min_col - 1),
        max(1, block_bounds.min_row - 2),
        block_bounds.max_col,
        min(block_bounds.max_row, block_bounds.min_row + 3),
    )
    records: list[dict[str, str]] = []
    seen: set[str] = set()
    for cell in _iter_range_cells(ws, scan_bounds):
        value = _cell_value(cell)
        if not value or value.startswith("=") or _is_company_specific(value) or _looks_numeric(value):
            continue
        if len(value) > 100:
            continue
        if value in seen:
            continue
        records.append(_cell_record(cell))
        seen.add(value)
        if len(records) >= 12:
            break
    if records:
        return records
    for label in anchor_labels:
        if label and not _is_company_specific(label):
            records.append({"cell": f"A{block_bounds.min_row}", "value": label})
            break
    return records


def _collect_static_label_cells(ws: Any, block_bounds: Bounds) -> list[dict[str, str]]:
    label_col = max(1, block_bounds.min_col - 1)
    records: list[dict[str, str]] = []
    seen: set[str] = set()
    for row_idx in range(block_bounds.min_row, min(block_bounds.max_row, ws.max_row) + 1):
        cell = ws.cell(row_idx, label_col)
        value = _cell_value(cell)
        if not value or value.startswith("=") or _is_company_specific(value) or _looks_numeric(value):
            continue
        if len(value) > 120:
            continue
        if value in seen:
            continue
        records.append(_cell_record(cell))
        seen.add(value)
        if len(records) >= 40:
            return records
    return records


def _count_nonempty_formula_numeric_text(ws: Any, target: str) -> dict[str, int]:
    target_bounds = _bounds(target)
    nonempty = 0
    formulas = 0
    numeric = 0
    text = 0
    for cell in _iter_range_cells(ws, target_bounds):
        value = cell.value
        if value is None or value == "":
            continue
        nonempty += 1
        if isinstance(value, str) and value.startswith("="):
            formulas += 1
        elif _looks_numeric(value):
            numeric += 1
        else:
            text += 1
    return {
        "nonempty_count": nonempty,
        "formula_count": formulas,
        "numeric_count": numeric,
        "text_count": text,
    }


def _sample_cells(ws: Any, target: str, predicate: Any, *, limit: int = 12) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    for cell in _iter_range_cells(ws, _bounds(target)):
        value = cell.value
        if value is None or value == "":
            continue
        if predicate(value):
            records.append(_cell_record(cell))
            if len(records) >= limit:
                break
    return records


def _writable_examples(ws: Any, bindings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    examples: list[dict[str, Any]] = []
    for entry in bindings:
        target = str(entry["target"])
        summary = _count_nonempty_formula_numeric_text(ws, target)
        examples.append(
            {
                "binding_id": entry["binding_id"],
                "target": target,
                "value_shape": entry["value_shape"],
                **summary,
                "sample_cells": _sample_cells(
                    ws,
                    target,
                    lambda value: not (isinstance(value, str) and value.startswith("=")),
                    limit=8,
                ),
            }
        )
    return examples


def _classified_examples(ws: Any, bindings: list[dict[str, Any]], predicate: Any) -> list[dict[str, str]]:
    seen: set[str] = set()
    records: list[dict[str, str]] = []
    for entry in bindings:
        for record in _sample_cells(ws, str(entry["target"]), predicate, limit=12):
            if record["cell"] in seen:
                continue
            records.append(record)
            seen.add(record["cell"])
            if len(records) >= 40:
                return records
    return records


def _collect_formula_cells(ws: Any, block_bounds: Bounds) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    for cell in _iter_range_cells(ws, block_bounds):
        if isinstance(cell.value, str) and cell.value.startswith("="):
            records.append(_formula_record(cell))
            if len(records) >= 50:
                break
    return records


def _collect_hidden_helpers(ws: Any, block_bounds: Bounds) -> list[dict[str, str]]:
    helpers: list[dict[str, str]] = []
    for col_idx in range(block_bounds.min_col, min(block_bounds.max_col, ws.max_column) + 1):
        letter = get_column_letter(col_idx)
        if ws.column_dimensions[letter].hidden:
            helpers.append({"type": "hidden_column", "target": letter})
    for row_idx in range(block_bounds.min_row, min(block_bounds.max_row, ws.max_row) + 1):
        if ws.row_dimensions[row_idx].hidden:
            helpers.append({"type": "hidden_row", "target": str(row_idx)})
    return helpers[:40]


def _merge_count(ws: Any, block_bounds: Bounds) -> int:
    return sum(
        1
        for merged in ws.merged_cells.ranges
        if _intersects(
            block_bounds,
            Bounds(merged.min_col, merged.min_row, merged.max_col, merged.max_row),
        )
    )


def _nonempty_and_formula_counts(ws: Any, block_bounds: Bounds) -> tuple[int, int]:
    nonempty = 0
    formulas = 0
    for cell in _iter_range_cells(ws, block_bounds):
        if cell.value is None:
            continue
        nonempty += 1
        if isinstance(cell.value, str) and cell.value.startswith("="):
            formulas += 1
    return nonempty, formulas


def _style_signature(ws: Any, block_bounds: Bounds) -> tuple[str, ...]:
    signatures: list[str] = []
    for cell in _iter_range_cells(ws, block_bounds):
        if cell.value is None:
            continue
        fill = _text(getattr(cell.fill.fgColor, "rgb", ""))
        font = "bold" if cell.font and cell.font.bold else "regular"
        border = "|".join(
            _text(getattr(side, "style", ""))
            for side in (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom)
        )
        signatures.append(f"{fill}:{font}:{border}")
        if len(signatures) >= 25:
            break
    return tuple(signatures)


def _average_row_height(ws: Any, block_bounds: Bounds) -> float:
    heights: list[float] = []
    for row_idx in range(block_bounds.min_row, min(block_bounds.max_row, ws.max_row) + 1):
        height = ws.row_dimensions[row_idx].height
        heights.append(float(height) if height is not None else 15.0)
    if not heights:
        return 0.0
    return sum(heights) / len(heights)


def _similarity(values: list[Any], *, numeric: bool = False) -> str:
    present = [value for value in values if value not in (None, "", "missing")]
    if len(present) < len(values):
        return "missing"
    if len(set(present)) == 1:
        return "same"
    if numeric:
        numbers = [float(value) for value in present]
        if min(numbers) == 0:
            return "similar" if max(numbers) <= 2 else "different"
        return "similar" if max(numbers) / max(min(numbers), 0.0001) <= 1.25 else "different"
    return "similar"


def _field_family(field: str) -> str:
    return field.split(".", 1)[0]


def _sheet_flow_by_name(sheet_flow: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {sheet["sheet"]: sheet for sheet in sheet_flow.get("sheets", [])}


def _support_sheets_for(sheet: str, flow: dict[str, Any]) -> list[str]:
    support: set[str] = set()
    sheet_flow = _sheet_flow_by_name(flow).get(sheet, {})
    storage = sheet_flow.get("storage_layer", {})
    support.update(storage.get("support_sheets_currently_store_intermediate_data", []))
    for dep in flow.get("support_sheet_dependencies", []):
        if sheet in dep.get("feeds_visible_sheets", []):
            support.add(dep["name"])
    return sorted(support)


def _owners_for(sheet: str, flow: dict[str, Any]) -> tuple[list[str], list[str]]:
    sheet_flow = _sheet_flow_by_name(flow).get(sheet, {})
    ownership = sheet_flow.get("ownership", {})
    return (
        list(ownership.get("current_code_owner_modules", [])),
        list(ownership.get("future_intended_owner_modules", [])),
    )


def _source_paths(source_dir: Path) -> dict[str, Path]:
    paths: dict[str, Path] = {}
    for ticker in SOURCE_TICKERS:
        macro_free = source_dir / f"{ticker}_model.xlsx"
        macro_enabled = source_dir / f"{ticker}_model.xlsm"
        paths[ticker] = macro_free if macro_free.exists() else macro_enabled
    return paths


def _copy_template_lab(source_dir: Path, lab_path: Path) -> dict[str, Any]:
    source_path = source_dir / "ANF_model.xlsx"
    lab_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source_path, lab_path)
    source_hash = _sha256(source_path)
    lab_hash = _sha256(lab_path)
    source_label, lab_label = _portable_template_lab_paths(
        source_path=source_path, source_dir=source_dir, lab_path=lab_path
    )
    return {
        "purpose": "Read-only-derived ANF workbook lab copy for template analysis; not canonical output.",
        "source_path": source_label,
        "lab_path": lab_label,
        "source_sha256": source_hash,
        "lab_sha256": lab_hash,
        "byte_identical": source_hash == lab_hash,
    }


def _existing_template_lab(source_dir: Path, lab_path: Path) -> dict[str, Any]:
    source_path = source_dir / "ANF_model.xlsx"
    if not lab_path.exists():
        raise FileNotFoundError(f"Existing template lab is required in read-only mode: {lab_path}")
    source_hash = _sha256(source_path)
    lab_hash = _sha256(lab_path)
    source_label, lab_label = _portable_template_lab_paths(
        source_path=source_path, source_dir=source_dir, lab_path=lab_path
    )
    return {
        "purpose": "Read-only-derived ANF workbook lab copy for template analysis; not canonical output.",
        "source_path": source_label,
        "lab_path": lab_label,
        "source_sha256": source_hash,
        "lab_sha256": lab_hash,
        "byte_identical": source_hash == lab_hash,
    }


def _source_workbook_meta(
    paths: dict[str, Path], *, source_dir: Path
) -> dict[str, dict[str, Any]]:
    meta: dict[str, dict[str, Any]] = {}
    for ticker, path in paths.items():
        meta[ticker] = {
            "path": f"@source_dir/{path.resolve().relative_to(source_dir.resolve()).as_posix()}",
            "sha256": _sha256(path) if path.exists() else "",
            "exists": path.exists(),
        }
    return meta


def _load_workbooks(paths: dict[str, Path]) -> dict[str, Any]:
    return {
        ticker: load_workbook(path, data_only=False, read_only=False)
        for ticker, path in paths.items()
        if path.exists()
    }


def _close_workbooks(workbooks: dict[str, Any]) -> None:
    for wb in workbooks.values():
        wb.close()


def _bindings_by_zone(bindings: list[dict[str, Any]]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for entry in bindings:
        if not entry.get("writable"):
            continue
        grouped.setdefault((entry["sheet"], entry["shell_zone"]), []).append(entry)
    return grouped


def _binding_records(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "binding_id": entry["binding_id"],
            "target": entry["target"],
            "normalized_field": entry["normalized_field"],
            "required": bool(entry["required"]),
            "value_shape": entry["value_shape"],
            "source_policy": entry["source_policy"],
            "missing_source_behavior": entry["missing_source_behavior"],
            "validation_rule": entry["validation_rule"],
        }
        for entry in entries
    ]


def _block_for_zone(
    *,
    sheet_name: str,
    zone: dict[str, Any],
    bindings: list[dict[str, Any]],
    sheet_bindings: list[dict[str, Any]],
    flow: dict[str, Any],
    anf_wb: Any,
) -> dict[str, Any]:
    block_bounds = _bounds(zone["target"])
    resolved_sheet = _sheet_name(sheet_name, "ANF")
    ws = anf_wb[resolved_sheet]
    anchor_labels = sorted({entry.get("anchor_label") or entry.get("section") or "" for entry in bindings})
    normalized_fields = sorted({entry["normalized_field"] for entry in bindings})
    required_fields = sorted({entry["normalized_field"] for entry in bindings if entry.get("required")})
    optional_fields = sorted({entry["normalized_field"] for entry in bindings if not entry.get("required")})
    source_policies = sorted({entry["source_policy"] for entry in bindings})
    validation_rules = sorted({entry["validation_rule"] for entry in bindings})
    missing_behaviors = sorted({entry["missing_source_behavior"] for entry in bindings})
    current_owners, future_owners = _owners_for(sheet_name, flow)

    template_label_cells = _outside_binding_targets(_collect_title_cells(ws, block_bounds, anchor_labels), sheet_bindings)
    row_label_cells = _outside_binding_targets(_collect_static_label_cells(ws, block_bounds), sheet_bindings)
    writable_example_cells = _writable_examples(ws, bindings)
    company_specific_example_cells = _classified_examples(
        ws,
        bindings,
        lambda value: _is_company_specific(_text(value)),
    )
    source_specific_example_cells = _classified_examples(
        ws,
        bindings,
        lambda value: isinstance(value, str)
        and not value.startswith("=")
        and _is_source_specific(value),
    )

    return {
        "block_id": zone["zone_id"],
        "sheet": sheet_name,
        "resolved_lab_sheet": resolved_sheet,
        "range": zone["target"],
        "title_header_cells": template_label_cells,
        "template_label_cells": template_label_cells,
        "row_label_cells": row_label_cells,
        "static_label_cells": [*template_label_cells, *row_label_cells],
        "writable_value_cells": [
            {
                "binding_id": entry["binding_id"],
                "target": entry["target"],
                "value_shape": entry["value_shape"],
            }
            for entry in bindings
        ],
        "writable_example_cells": writable_example_cells,
        "company_specific_example_cells": company_specific_example_cells,
        "source_specific_example_cells": source_specific_example_cells,
        "formula_cells": _collect_formula_cells(ws, block_bounds),
        "hidden_helper_cells": _collect_hidden_helpers(ws, block_bounds),
        "normalized_fields": normalized_fields,
        "support_sheets_used": _support_sheets_for(sheet_name, flow),
        "current_code_owner": current_owners,
        "future_intended_owner": future_owners,
        "required_fields": required_fields,
        "optional_fields": optional_fields,
        "source_policy": source_policies[0] if len(source_policies) == 1 else "mixed",
        "missing_data_behavior": " | ".join(missing_behaviors),
        "validation_rules": validation_rules,
        "standardization_status": "standard",
        "bindings": _binding_records(bindings),
        "notes": [
            "Block is derived from manifest writable zone and current binding map.",
            "ANF/PBI/GPRE workbook content is read-only evidence; future shell must clear writable company-specific values.",
        ],
    }


def _coverage_status(wb: Any, ticker: str, sheet_name: str, range_ref: str) -> dict[str, Any]:
    resolved_sheet = _sheet_name(sheet_name, ticker)
    if resolved_sheet not in wb.sheetnames:
        return {
            "block_exists": False,
            "resolved_sheet": resolved_sheet,
            "range": range_ref,
            "nonempty_cells": 0,
            "formula_cells": 0,
            "merge_count": 0,
            "freeze_panes": "",
            "max_row": 0,
            "max_column": 0,
            "style_signature": [],
            "average_row_height": 0.0,
        }
    ws = wb[resolved_sheet]
    block_bounds = _bounds(range_ref)
    nonempty, formulas = _nonempty_and_formula_counts(ws, block_bounds)
    exists = ws.max_row >= block_bounds.min_row and ws.max_column >= block_bounds.min_col
    return {
        "block_exists": bool(exists),
        "resolved_sheet": resolved_sheet,
        "range": range_ref,
        "nonempty_cells": nonempty,
        "formula_cells": formulas,
        "merge_count": _merge_count(ws, block_bounds),
        "freeze_panes": str(ws.freeze_panes or ""),
        "max_row": int(ws.max_row),
        "max_column": int(ws.max_column),
        "style_signature": list(_style_signature(ws, block_bounds)),
        "average_row_height": round(_average_row_height(ws, block_bounds), 2),
    }


def _coverage_row(block: dict[str, Any], workbooks: dict[str, Any]) -> dict[str, Any]:
    statuses = {
        ticker: _coverage_status(workbooks[ticker], ticker, block["sheet"], block["range"])
        for ticker in SOURCE_TICKERS
    }
    merge_values = [statuses[ticker]["merge_count"] for ticker in SOURCE_TICKERS]
    freeze_values = [statuses[ticker]["freeze_panes"] or "missing" for ticker in SOURCE_TICKERS]
    row_height_values = [statuses[ticker]["average_row_height"] for ticker in SOURCE_TICKERS]
    style_values = [tuple(statuses[ticker]["style_signature"]) for ticker in SOURCE_TICKERS]
    families = sorted({_field_family(field) for field in block["normalized_fields"]})
    differences: list[str] = []
    for ticker in SOURCE_TICKERS:
        status = statuses[ticker]
        if not status["block_exists"]:
            differences.append(f"{ticker}: missing block or sheet")
        if status["max_row"] and status["max_row"] < _bounds(block["range"]).max_row:
            differences.append(f"{ticker}: source sheet shorter than block range")
    if block["sheet"] == "Operating_Drivers":
        differences.append("GPRE has separate sector overlays that are excluded from standard blocks")

    return {
        "block_id": block["block_id"],
        "sheet": block["sheet"],
        "ANF": {key: value for key, value in statuses["ANF"].items() if key != "style_signature"},
        "PBI": {key: value for key, value in statuses["PBI"].items() if key != "style_signature"},
        "GPRE": {key: value for key, value in statuses["GPRE"].items() if key != "style_signature"},
        "range_similarity": "same" if all(status["block_exists"] for status in statuses.values()) else "missing",
        "style_similarity": "same" if len(set(style_values)) == 1 else "similar",
        "merge_similarity": _similarity(merge_values, numeric=True),
        "freeze_pane_similarity": "same" if len(set(freeze_values)) == 1 else "different",
        "row_height_similarity": _similarity(row_height_values, numeric=True),
        "populated_field_families": families,
        "support_sheet_dependencies": block["support_sheets_used"],
        "ticker_specific_differences": differences,
        "include_in_standard_template": block["standardization_status"] == "standard",
        "exclusion_reason": "",
    }


def _excluded_sector_overlays(workbooks: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet in SECTOR_OVERLAYS:
        exists_in = [ticker for ticker, wb in workbooks.items() if sheet in wb.sheetnames]
        rows.append(
            {
                "sheet": sheet,
                "exists_in": exists_in,
                "standardization_status": "sector_specific",
                "include_in_standard_template": False,
                "exclusion_reason": "GPRE-specific sector overlay; not part of standard visible template family.",
            }
        )
    return rows


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def _fmt_list(values: Iterable[str], *, limit: int = 4) -> str:
    items = [str(value) for value in values if str(value)]
    if not items:
        return "-"
    if len(items) > limit:
        return ", ".join(items[:limit]) + f", +{len(items) - limit} more"
    return ", ".join(items)


def _write_architecture_md(path: Path, payload: dict[str, Any]) -> None:
    lines = [
        "# Workbook Block Architecture",
        "",
        "This is a read-only block map for the future rich standard template shell. It is generated from the ANF lab workbook, PBI/GPRE cross-check workbooks, the shell manifest, the binding map, and the sheet data-flow map.",
        "",
        "It is not a runtime filler and it must not be used to patch or promote ticker workbooks.",
        "",
        "## Template Lab",
        "",
        f"- Lab workbook: `{payload['template_lab']['lab_path']}`",
        f"- Source workbook: `{payload['template_lab']['source_path']}`",
        f"- Byte-identical copy: `{payload['template_lab']['byte_identical']}`",
        f"- Source SHA256: `{payload['template_lab']['source_sha256']}`",
        "",
        "## Standard vs Optional Sector Packs",
        "",
        "The standard shell keeps generic block slots only. Sector/company member names from ANF/PBI/GPRE are evidence for optional packs or clearing, not standard template labels.",
        "",
        "| Pack | Status | Example members | Activation rule |",
        "| --- | --- | --- | --- |",
    ]
    for pack in payload.get("optional_sector_packs", []):
        lines.append(
            "| {name} | {status} | {members} | {rule} |".format(
                name=pack["display_name"],
                status=pack["status"],
                members=_fmt_list(pack["example_members"], limit=8),
                rule=pack["activation_rule"],
            )
        )
    lines.extend(
        [
            "",
        "## Block Summary",
        "",
        "| Block | Sheet | Range | Fields | Policy | Standardization |",
        "| --- | --- | --- | --- | --- | --- |",
        ]
    )
    for block in payload["blocks"]:
        lines.append(
            "| {block_id} | {sheet} | {range_ref} | {fields} | {policy} | {status} |".format(
                block_id=block["block_id"],
                sheet=block["sheet"],
                range_ref=block["range"],
                fields=_fmt_list(block["normalized_fields"]),
                policy=block["source_policy"],
                status=block["standardization_status"],
            )
        )
    lines.extend(["", "## Sheet Blocks", ""])
    for sheet in payload["standard_visible_sheets"]:
        lines.append(f"### {sheet}")
        lines.append("")
        for block in [entry for entry in payload["blocks"] if entry["sheet"] == sheet]:
            lines.extend(
                [
                    f"- `{block['block_id']}` `{block['range']}`",
                    f"  - Normalized fields: {_fmt_list(block['normalized_fields'], limit=8)}",
                    f"  - Support sheets: {_fmt_list(block['support_sheets_used'], limit=8)}",
                    f"  - Current owner: {_fmt_list(block['current_code_owner'], limit=4)}",
                    f"  - Future owner: {_fmt_list(block['future_intended_owner'], limit=4)}",
                    f"  - Missing data: {block['missing_data_behavior']}",
                    f"  - Validation: {_fmt_list(block['validation_rules'], limit=8)}",
                    "",
                ]
            )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def _write_coverage_md(path: Path, payload: dict[str, Any]) -> None:
    lines = [
        "# Workbook Block Coverage Matrix",
        "",
        "This matrix compares the standard block map across ANF, PBI, and GPRE. ANF is the visual lab base; PBI and GPRE are cross-checks. GPRE-only sector overlays are explicitly excluded from standard-template behavior.",
        "",
        "## Coverage Rows",
        "",
        "| Block | Sheet | Include | Range | Style | Merge | Freeze | Differences |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in payload["coverage_rows"]:
        lines.append(
            "| {block_id} | {sheet} | {include} | {range_sim} | {style} | {merge} | {freeze} | {diffs} |".format(
                block_id=row["block_id"],
                sheet=row["sheet"],
                include="yes" if row["include_in_standard_template"] else "no",
                range_sim=row["range_similarity"],
                style=row["style_similarity"],
                merge=row["merge_similarity"],
                freeze=row["freeze_pane_similarity"],
                diffs=_fmt_list(row["ticker_specific_differences"], limit=3),
            )
        )
    lines.extend(["", "## Excluded Sector Overlays", ""])
    for entry in payload.get("excluded_sector_overlays", []):
        lines.append(
            f"- `{entry['sheet']}`: excluded because {entry['exclusion_reason']} Exists in: {_fmt_list(entry['exists_in'])}."
        )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def build(
    *,
    source_dir: Path,
    lab_path: Path,
    reuse_existing_lab: bool = False,
    architecture_json: Path = ARCHITECTURE_JSON,
    architecture_md: Path = ARCHITECTURE_MD,
    coverage_json: Path = COVERAGE_JSON,
    coverage_md: Path = COVERAGE_MD,
) -> tuple[dict[str, Any], dict[str, Any]]:
    manifest = _load_json(ROOT / "docs" / "standard_template_shell_manifest.json")
    binding_payload = _load_json(ROOT / "docs" / "workbook_binding_map.json")
    flow = _load_json(ROOT / "docs" / "sheet_data_flow_map.json")
    source_paths = _source_paths(source_dir)
    template_lab = _existing_template_lab(source_dir, lab_path) if reuse_existing_lab else _copy_template_lab(source_dir, lab_path)
    workbooks = _load_workbooks(source_paths)
    try:
        grouped_bindings = _bindings_by_zone(binding_payload["bindings"])
        blocks: list[dict[str, Any]] = []
        for sheet_def in manifest["sheets"]:
            sheet_name = sheet_def["sheet"]
            if sheet_name not in STANDARD_VISIBLE_SHEETS:
                continue
            sheet_bindings = [entry for entry in binding_payload["bindings"] if entry["sheet"] == sheet_name]
            for zone in sheet_def["writable_zones"]:
                bindings = grouped_bindings.get((sheet_name, zone["zone_id"]), [])
                if not bindings:
                    continue
                blocks.append(
                    _block_for_zone(
                        sheet_name=sheet_name,
                        zone=zone,
                        bindings=bindings,
                        sheet_bindings=sheet_bindings,
                        flow=flow,
                        anf_wb=workbooks["ANF"],
                    )
                )

        generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        architecture = {
            "version": "0.1.0",
            "generated_at": generated_at,
            "source_workbooks": _source_workbook_meta(
                source_paths, source_dir=source_dir
            ),
            "template_lab": template_lab,
            "standard_visible_sheets": STANDARD_VISIBLE_SHEETS,
            "optional_sector_packs": OPTIONAL_SECTOR_PACKS,
            "standard_shell_neutrality_rule": "Standard visible blocks use generic slots; fixed company, sector, and dimension members belong in optional sector packs or normalized data.",
            "blocks": blocks,
        }
        coverage = {
            "version": "0.1.0",
            "generated_at": generated_at,
            "source_workbooks": _source_workbook_meta(
                source_paths, source_dir=source_dir
            ),
            "coverage_rows": [_coverage_row(block, workbooks) for block in blocks],
            "excluded_sector_overlays": _excluded_sector_overlays(workbooks),
        }
    finally:
        _close_workbooks(workbooks)

    _write_json(architecture_json, architecture)
    _write_json(coverage_json, coverage)
    _write_architecture_md(architecture_md, architecture)
    _write_coverage_md(coverage_md, coverage)
    return architecture, coverage


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-dir", type=Path, default=DEFAULT_SOURCE_DIR)
    parser.add_argument("--lab-path", type=Path, default=DEFAULT_LAB_PATH)
    parser.add_argument("--reuse-existing-lab", action="store_true", help="Read the existing lab workbook without copying or modifying any xlsx file.")
    parser.add_argument("--architecture-json", type=Path, default=ARCHITECTURE_JSON)
    parser.add_argument("--architecture-md", type=Path, default=ARCHITECTURE_MD)
    parser.add_argument("--coverage-json", type=Path, default=COVERAGE_JSON)
    parser.add_argument("--coverage-md", type=Path, default=COVERAGE_MD)
    args = parser.parse_args(argv)

    is_default_run = (
        args.source_dir.expanduser().resolve() == DEFAULT_SOURCE_DIR.resolve()
        and args.lab_path.expanduser().resolve() == DEFAULT_LAB_PATH.resolve()
        and args.architecture_json.expanduser().resolve() == ARCHITECTURE_JSON.resolve()
        and args.architecture_md.expanduser().resolve() == ARCHITECTURE_MD.resolve()
        and args.coverage_json.expanduser().resolve() == COVERAGE_JSON.resolve()
        and args.coverage_md.expanduser().resolve() == COVERAGE_MD.resolve()
    )
    if is_default_run and os.environ.get("STANDARD_TEMPLATE_AUDIT_ISOLATED_RUN") != "1":
        run_audit_generator(Path(__file__), root=ROOT, data_root=DEFAULT_DATA_ROOT)
        architecture = _load_json(ARCHITECTURE_JSON)
        coverage = _load_json(COVERAGE_JSON)
    else:
        architecture, coverage = build(
            source_dir=args.source_dir.expanduser().resolve(),
            lab_path=args.lab_path.expanduser().resolve(),
            reuse_existing_lab=args.reuse_existing_lab,
            architecture_json=args.architecture_json.expanduser().resolve(),
            architecture_md=args.architecture_md.expanduser().resolve(),
            coverage_json=args.coverage_json.expanduser().resolve(),
            coverage_md=args.coverage_md.expanduser().resolve(),
        )
    print(f"template lab: {architecture['template_lab']['lab_path']}")
    print(f"blocks: {len(architecture['blocks'])}")
    print(f"coverage rows: {len(coverage['coverage_rows'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
