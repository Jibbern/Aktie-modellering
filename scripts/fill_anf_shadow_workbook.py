"""Render an ANF shadow workbook through the value-only new-ticker runtime.

This script is intentionally ANF-shadow-only. It reads the cleaned normalized
package and frozen shell, validates a temporary candidate, atomically promotes
ANF_shadow_model.xlsx only after PASS, and emits coverage/audit reports. It does
not call production workbook writers or replace ANF_model.xlsx.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import range_boundaries

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from pbi_xbrl.json_schema_validation import load_json_strict  # noqa: E402
from pbi_xbrl.new_ticker_value_filler import (  # noqa: E402
    DEFAULT_BINDING_MAP,
    DEFAULT_MANIFEST,
    DEFAULT_TEMPLATE,
    fill_standard_template_from_package,
)
from pbi_xbrl.new_ticker_binding_planner import (  # noqa: E402
    BindingPlan,
    reproduce_binding_plan,
    write_binding_plan_report,
)
from pbi_xbrl.normalized_company_data_validation import (  # noqa: E402
    build_mapping_gap_report,
    validate_normalized_company_data,
)
from scripts.validate_standard_template_shell import validate_shell  # noqa: E402


MINIMUM_USEFULNESS = {
    "quarterly_financial_rows": 8,
    "annual_financial_rows": 3,
    "guidance_rows": 5,
    "segment_rows": 5,
    "operating_driver_visible_rows": 5,
    "quarter_note_visible_rows": 8,
}
STANDARD_VISIBLE_SHEETS = [
    "SUMMARY",
    "Valuation",
    "BS_Segments",
    "Operating_Drivers",
    "{ticker}_Investment_Case",
    "Quarter_Notes_UI",
    "Promise_Progress_UI",
    "QA_Log",
    "Needs_Review",
    "QA_Checks",
]


def _default_data_root() -> Path:
    for ancestor in [REPO_ROOT, *REPO_ROOT.parents]:
        candidate = ancestor / "StockModelData"
        if candidate.exists():
            return candidate
    return REPO_ROOT.parent / "StockModelData"


def _load_json(path: Path) -> dict[str, Any]:
    payload = load_json_strict(path)
    if not isinstance(payload, dict):
        raise ValueError(f"JSON contract must be an object: {path}")
    return payload


def _write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _create_candidate_workbook(final_path: Path) -> Path:
    final_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, raw_path = tempfile.mkstemp(
        prefix=f".{final_path.stem}.",
        suffix=".candidate.xlsx",
        dir=final_path.parent,
    )
    os.close(descriptor)
    return Path(raw_path)


def _atomic_promote_workbook(candidate_path: Path, final_path: Path) -> None:
    if candidate_path.parent.resolve() != final_path.parent.resolve():
        raise RuntimeError("Shadow candidate and final workbook must share a directory for atomic promotion.")
    os.replace(candidate_path, final_path)


def _resolve_sheet(sheet_name: str, ticker: str = "ANF") -> str:
    return sheet_name.replace("{ticker}", ticker)


def _field_value(value: Any) -> Any:
    if isinstance(value, Mapping):
        if str(value.get("status") or "") != "populated":
            return None
        return value.get("value")
    return value


def _path_get(obj: Any, dotted_path: str) -> Any:
    current = obj
    for part in dotted_path.split("."):
        if isinstance(current, Mapping):
            if part not in current:
                return None
            current = current[part]
            continue
        if isinstance(current, list):
            try:
                current = current[int(part)]
            except (IndexError, ValueError):
                return None
            continue
        return None
    return current


def _section_populated_count(obj: Any) -> int:
    if isinstance(obj, Mapping):
        if "value" in obj and str(obj.get("status") or "") == "populated" and obj.get("value") not in (None, ""):
            return 1
        return sum(_section_populated_count(value) for value in obj.values())
    if isinstance(obj, list):
        return sum(_section_populated_count(value) for value in obj)
    return 0


def _is_populated_field(value: Any) -> bool:
    return isinstance(value, Mapping) and str(value.get("status") or "") == "populated" and value.get("value") not in (None, "")


def _visible_row_counts(package: Mapping[str, Any]) -> dict[str, int]:
    guidance = package.get("normalized_guidance", {}).get("items", [])
    segments = package.get("segments", {}).get("items", [])
    drivers = package.get("operating_drivers", {}).get("items", [])
    notes = package.get("quarter_notes", {}).get("items", [])
    return {
        "summary_fields": _section_populated_count(package.get("company_profile", {}))
        + _section_populated_count(package.get("investment_case", {})),
        "valuation_financial_rows": len(package.get("quarterly_financials", {}).get("rows", [])),
        "quarterly_financial_rows": len(package.get("quarterly_financials", {}).get("rows", [])),
        "annual_financial_rows": len(package.get("annual_financials", {}).get("rows", [])),
        "guidance_rows": len([item for item in guidance if _is_populated_field(item.get("value"))]),
        "segment_rows": len(segments),
        "operating_driver_visible_rows": len([item for item in drivers if _is_populated_field(item.get("current_read"))]),
        "quarter_note_visible_rows": len([item for item in notes if _is_populated_field(item.get("note")) or _is_populated_field(item.get("commentary"))]),
        "investment_case_fields": _section_populated_count(package.get("investment_case", {})),
        "qa_log_rows": len(package.get("manual_review_flags", [])),
        "needs_review_rows": len(package.get("manual_review_flags", [])),
        "qa_checks_rows": len(package.get("mapping_gaps", [])),
    }


def _manual_review_summary(package: Mapping[str, Any]) -> dict[str, Any]:
    flags = [item for item in package.get("manual_review_flags", []) if isinstance(item, Mapping)]
    return {
        "total_count": len(flags),
        "by_severity": dict(Counter(str(item.get("severity") or "") for item in flags)),
        "by_rule_id": dict(Counter(str(item.get("rule_id") or "") for item in flags)),
        "by_classification": dict(Counter(str(item.get("classification") or "") for item in flags if item.get("classification"))),
    }


def _binding_preview(plan: BindingPlan, bindings: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    by_sheet: dict[str, dict[str, Any]] = {}
    rows: list[dict[str, Any]] = []
    writes_by_binding: dict[str, list[Any]] = defaultdict(list)
    for write in plan.planned_writes:
        writes_by_binding[write.binding_id].append(write)
    for binding in bindings:
        if not bool(binding.get("writable")) or str(binding.get("source_policy") or "") == "validation-output":
            continue
        binding_writes = writes_by_binding.get(str(binding.get("binding_id") or ""), [])
        cell_count = len(binding_writes)
        row_count = len({write.row_key for write in binding_writes})
        sheet = str(binding.get("sheet") or "")
        row = {
            "binding_id": str(binding.get("binding_id") or ""),
            "sheet": sheet,
            "target": str(binding.get("target") or ""),
            "normalized_field": str(binding.get("normalized_field") or ""),
            "required": bool(binding.get("required")),
            "value_shape": str(binding.get("value_shape") or ""),
            "available_rows": row_count,
            "available_cells": cell_count,
            "would_write_useful_output": cell_count > 0,
        }
        rows.append(row)
        summary = by_sheet.setdefault(
            sheet,
            {"binding_count": 0, "bindings_with_data": 0, "available_rows": 0, "available_cells": 0},
        )
        summary["binding_count"] += 1
        summary["available_rows"] += row_count
        summary["available_cells"] += cell_count
        if cell_count:
            summary["bindings_with_data"] += 1
    return {"by_binding": rows, "by_sheet": by_sheet}


def build_prefill_coverage_report(
    package: Mapping[str, Any],
    bindings: Sequence[Mapping[str, Any]],
    *,
    plan: BindingPlan,
) -> dict[str, Any]:
    validation_issues = validate_normalized_company_data(package, binding_map=bindings, promotion_requested=False)
    p1_issues = [issue.to_dict() for issue in validation_issues if issue.severity.upper() in {"P0", "P1"}]
    raw_mapping = build_mapping_gap_report(package, bindings, ticker="ANF")
    mapping_gaps = [
        gap
        for gap in raw_mapping.get("gaps", [])
        if str(gap.get("source_policy") or "") != "validation-output"
    ]
    visible_rows = _visible_row_counts(package)
    minimum_issues: list[dict[str, Any]] = []
    for metric, minimum in MINIMUM_USEFULNESS.items():
        actual = int(visible_rows.get(metric, 0))
        if actual < minimum:
            minimum_issues.append(
                {
                    "severity": "P2",
                    "rule_id": "anf_shadow_usefulness_observation",
                    "field": metric,
                    "message": f"ANF shadow package has {actual} useful rows; suggested comparison minimum is {minimum}.",
                    "suggested_action": "Consider improving source-backed coverage before promotion; do not backfill generic text.",
                }
            )
    source_coverage = package.get("source_coverage", {}) if isinstance(package.get("source_coverage"), Mapping) else {}
    return {
        "ticker": "ANF",
        "status": "FAIL" if p1_issues else "PASS",
        "populated_fields_by_section": {
            section: _section_populated_count(package.get(section))
            for section in (
                "ticker_metadata",
                "company_profile",
                "quarterly_financials",
                "annual_financials",
                "debt_liquidity",
                "capital_returns",
                "normalized_guidance",
                "segments",
                "operating_drivers",
                "quarter_notes",
                "investment_case",
            )
        },
        "visible_rows_available": visible_rows,
        "demoted_rows": source_coverage.get("text_quality_summary", {"total_demoted": 0}),
        "manual_review_flags": _manual_review_summary(package),
        "mapping_gaps": {
            "total_count": len(mapping_gaps),
            "by_binding_id": dict(Counter(str(gap.get("binding_id") or "") for gap in mapping_gaps)),
            "gaps": mapping_gaps,
        },
        "expected_useful_output_by_visible_sheet": _binding_preview(plan, bindings)["by_sheet"],
        "binding_preview": _binding_preview(plan, bindings)["by_binding"],
        "minimum_usefulness": {
            "status": "WARN" if minimum_issues else "PASS",
            "minimums": MINIMUM_USEFULNESS,
            "issues": minimum_issues,
        },
        "normalized_validation": {
            "issue_count": len(validation_issues),
            "p1_issue_count": len(p1_issues),
            "p1_issues": p1_issues,
        },
    }


def _markdown_prefill(report: Mapping[str, Any]) -> str:
    lines = [
        "ANF shadow prefill coverage",
        f"Status: {report['status']}",
        "",
        "Visible rows available:",
    ]
    for key, value in report["visible_rows_available"].items():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "Demoted rows:"])
    demoted = report.get("demoted_rows", {})
    lines.append(f"- total_demoted: {demoted.get('total_demoted', 0)}")
    for key, value in (demoted.get("by_classification", {}) or {}).items():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "Expected useful output by sheet:"])
    for sheet, row in report["expected_useful_output_by_visible_sheet"].items():
        lines.append(f"- {sheet}: {row['available_rows']} rows / {row['available_cells']} cells")
    if report["minimum_usefulness"]["issues"]:
        lines.extend(["", "Usefulness observations:"])
        for issue in report["minimum_usefulness"]["issues"]:
            lines.append(f"- {issue['field']}: {issue['message']}")
    return "\n".join(lines) + "\n"


def _writable_ranges_by_sheet(manifest: Mapping[str, Any], ticker: str = "ANF") -> dict[str, list[str]]:
    return {
        _resolve_sheet(str(sheet["sheet"]), ticker): [str(zone["target"]) for zone in sheet.get("writable_zones", [])]
        for sheet in manifest.get("sheets", [])
    }


def _written_cells_by_sheet(template_path: Path, output_path: Path, manifest: Mapping[str, Any], ticker: str = "ANF") -> dict[str, dict[str, int]]:
    writable_ranges = _writable_ranges_by_sheet(manifest, ticker)
    before = load_workbook(template_path, data_only=False, read_only=False)
    after = load_workbook(output_path, data_only=False, read_only=False)
    try:
        summary: dict[str, dict[str, int]] = {}
        for template_ws in before.worksheets:
            sheet_name = _resolve_sheet(template_ws.title, ticker)
            if sheet_name not in after.sheetnames:
                continue
            output_ws = after[sheet_name]
            written_rows: set[int] = set()
            count = 0
            for range_ref in writable_ranges.get(sheet_name, []):
                min_col, min_row, max_col, max_row = range_boundaries(range_ref)
                for row_idx in range(min_row, max_row + 1):
                    for col_idx in range(min_col, max_col + 1):
                        before_cell = template_ws.cell(row_idx, col_idx)
                        after_cell = output_ws.cell(row_idx, col_idx)
                        if isinstance(before_cell, MergedCell) or isinstance(after_cell, MergedCell):
                            continue
                        if after_cell.value not in (None, "") and before_cell.value != after_cell.value:
                            count += 1
                            written_rows.add(row_idx)
            summary[sheet_name] = {"written_cell_count": count, "written_row_count": len(written_rows)}
        return summary
    finally:
        before.close()
        after.close()


def _count_non_empty_rows(path: Path, sheet_name: str, *, min_row: int = 2, max_rows: int = 5000) -> int:
    wb = load_workbook(path, data_only=False, read_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            return 0
        ws = wb[sheet_name]
        count = 0
        for row in ws.iter_rows(min_row=min_row, max_row=min(ws.max_row, max_rows)):
            if any(cell.value not in (None, "") for cell in row):
                count += 1
        return count
    finally:
        wb.close()


def build_post_fill_audit(
    *,
    template_path: Path,
    output_path: Path,
    manifest: Mapping[str, Any],
    bindings: Sequence[Mapping[str, Any]],
    package: Mapping[str, Any],
    plan: BindingPlan,
    strict_validation: Mapping[str, Any],
) -> dict[str, Any]:
    written = _written_cells_by_sheet(template_path, output_path, manifest)
    manual_rows = _count_non_empty_rows(output_path, "Needs_Review")
    qa_rows = _count_non_empty_rows(output_path, "QA_Log")
    gap_rows = _count_non_empty_rows(output_path, "QA_Checks")
    binding_reports = {
        str(report.get("binding_id") or ""): report
        for report in plan.binding_reports
        if isinstance(report, Mapping)
    }
    blank_required = [
        {
            "binding_id": str(binding.get("binding_id") or ""),
            "sheet": str(binding.get("sheet") or ""),
            "target": str(binding.get("target") or ""),
            "normalized_field": str(binding.get("normalized_field") or ""),
        }
        for binding in bindings
        if bool(binding.get("required"))
        and bool(binding.get("writable"))
        and str(binding.get("source_policy") or "") != "validation-output"
        and int(binding_reports.get(str(binding.get("binding_id") or ""), {}).get("planned_write_count") or 0) == 0
    ]
    identity = strict_validation.get("shell_identity") if isinstance(strict_validation.get("shell_identity"), Mapping) else {}
    strict_issues = [issue for issue in strict_validation.get("issues") or [] if isinstance(issue, Mapping)]
    strict_rule_ids = {str(issue.get("rule_id") or "") for issue in strict_issues}
    structural_rule_ids = {
        "post_fill_sheet_order_visibility_drift",
        "post_fill_merge_drift",
        "post_fill_defined_name_drift",
        "post_fill_layout_drift",
        "post_fill_protected_cell_drift",
        "post_fill_data_validation_drift",
        "post_fill_conditional_formatting_drift",
        "post_fill_table_drift",
    }
    visible_summary = {
        sheet: {
            **written.get(_resolve_sheet(sheet), {"written_cell_count": 0, "written_row_count": 0}),
            "blank_required_bindings": sum(1 for item in blank_required if item["sheet"] == sheet),
            "mapping_gaps_rendered": gap_rows if sheet == "QA_Checks" else 0,
            "manual_review_rows_rendered": manual_rows if sheet == "Needs_Review" else qa_rows if sheet == "QA_Log" else 0,
        }
        for sheet in STANDARD_VISIBLE_SHEETS
    }
    return {
        "ticker": "ANF",
        "status": str(strict_validation.get("status") or "FAIL"),
        "output_path": str(output_path),
        "approved_plan_status": plan.status,
        "approved_plan_write_count": len(plan.planned_writes),
        "written_cell_count": sum(row["written_cell_count"] for row in written.values()),
        "written_row_count": sum(row["written_row_count"] for row in written.values()),
        "visible_usefulness_by_sheet": visible_summary,
        "blank_required_bindings": blank_required,
        "mapping_gaps_rendered": gap_rows,
        "manual_review_rows_rendered": manual_rows,
        "qa_log_rows_rendered": qa_rows,
        "manual_review_rows_kept_json_only": max(0, len(package.get("manual_review_flags", [])) - manual_rows),
        "formulas_unchanged": "post_fill_protected_cell_drift" not in strict_rule_ids,
        "layout_signature_unchanged": not bool(strict_rule_ids & structural_rule_ids),
        "non_writable_cells_unchanged": "post_fill_protected_cell_drift" not in strict_rule_ids,
        "non_writable_value_diffs": [
            str(issue.get("message") or "")
            for issue in strict_issues
            if str(issue.get("rule_id") or "") == "post_fill_protected_cell_drift"
        ],
        "strict_post_fill_validation": {
            "status": str(strict_validation.get("status") or "FAIL"),
            "issue_count": int(strict_validation.get("issue_count") or len(strict_issues)),
            "issues": strict_issues,
            "changed_writable_cell_count": int(identity.get("changed_writable_cell_count") or 0),
        },
    }


def _markdown_postfill(report: Mapping[str, Any]) -> str:
    lines = [
        "ANF shadow workbook fill audit",
        f"Status: {report['status']}",
        f"Workbook: {report['output_path']}",
        f"Written cells: {report['written_cell_count']}",
        f"Written rows: {report['written_row_count']}",
        f"Formulas unchanged: {report['formulas_unchanged']}",
        f"Layout signature unchanged: {report['layout_signature_unchanged']}",
        f"Non-writable cells unchanged: {report['non_writable_cells_unchanged']}",
        "",
        "Visible usefulness by sheet:",
    ]
    for sheet, row in report["visible_usefulness_by_sheet"].items():
        lines.append(f"- {sheet}: {row['written_row_count']} rows / {row['written_cell_count']} cells")
    return "\n".join(lines) + "\n"


def _non_empty_count(ws: Any, range_ref: str) -> int:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    count = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value not in (None, ""):
                count += 1
    return count


def _changed_count(template_ws: Any, shadow_ws: Any, range_ref: str) -> int:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    count = 0
    for row_idx in range(min_row, max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            before = template_ws.cell(row_idx, col_idx)
            after = shadow_ws.cell(row_idx, col_idx)
            if isinstance(before, MergedCell) or isinstance(after, MergedCell):
                continue
            if after.value not in (None, "") and before.value != after.value:
                count += 1
    return count


def build_shadow_vs_legacy_block_comparison(
    *,
    template_path: Path,
    shadow_path: Path,
    legacy_path: Path,
    block_architecture_path: Path = REPO_ROOT / "docs" / "workbook_block_architecture.json",
    prefill_report: Mapping[str, Any],
) -> dict[str, Any]:
    architecture = _load_json(block_architecture_path) if block_architecture_path.exists() else {"blocks": []}
    template = load_workbook(template_path, data_only=False, read_only=False)
    shadow = load_workbook(shadow_path, data_only=False, read_only=False)
    legacy = load_workbook(legacy_path, data_only=False, read_only=True)
    rows: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    try:
        for block in architecture.get("blocks", []):
            sheet = str(block.get("sheet") or "")
            if sheet not in STANDARD_VISIBLE_SHEETS:
                continue
            range_ref = str(block.get("range") or "")
            if not range_ref:
                continue
            resolved_sheet = _resolve_sheet(sheet)
            template_sheet = sheet
            exists_both = resolved_sheet in shadow.sheetnames and resolved_sheet in legacy.sheetnames
            shadow_non_empty = 0
            legacy_non_empty = 0
            shadow_written = 0
            if exists_both and template_sheet in template.sheetnames:
                shadow_non_empty = _non_empty_count(shadow[resolved_sheet], range_ref)
                legacy_non_empty = _non_empty_count(legacy[resolved_sheet], range_ref)
                shadow_written = _changed_count(template[template_sheet], shadow[resolved_sheet], range_ref)
            row = {
                "block_id": str(block.get("block_id") or ""),
                "sheet": sheet,
                "range": range_ref,
                "exists_in_shadow_and_legacy": exists_both,
                "legacy_non_empty_cells": legacy_non_empty,
                "shadow_non_empty_cells": shadow_non_empty,
                "shadow_written_cells": shadow_written,
                "shadow_populated": shadow_written > 0,
                "standardization_status": str(block.get("standardization_status") or ""),
            }
            rows.append(row)
            if exists_both and legacy_non_empty > 0 and shadow_written == 0 and row["standardization_status"] == "standard":
                missing.append(row)
        return {
            "ticker": "ANF",
            "summary": {
                "blocks_compared": len(rows),
                "blocks_existing_in_both": sum(1 for row in rows if row["exists_in_shadow_and_legacy"]),
                "shadow_populated_blocks": sum(1 for row in rows if row["shadow_populated"]),
                "legacy_blocks_not_yet_covered": len(missing),
            },
            "blocks": rows,
            "legacy_blocks_not_yet_covered_by_shadow": missing[:50],
            "important_legacy_fields_missing_in_shadow": missing[:20],
            "shadow_fields_cleaner_or_more_traceable": [
                "Row-schema tables preserve source_ref/source columns for Promise_Progress_UI, Quarter_Notes_UI, Operating_Drivers and QA sheets.",
                "Demoted noisy legacy/parser text is visible in QA/Needs_Review rather than mixed into operating UI fields.",
            ],
            "top_binding_gaps_to_fix_next": prefill_report.get("mapping_gaps", {}).get("gaps", [])[:20],
        }
    finally:
        template.close()
        shadow.close()
        legacy.close()


def _markdown_comparison(report: Mapping[str, Any]) -> str:
    summary = report["summary"]
    lines = [
        "ANF shadow vs legacy block comparison",
        f"Blocks compared: {summary['blocks_compared']}",
        f"Blocks existing in both: {summary['blocks_existing_in_both']}",
        f"Shadow populated blocks: {summary['shadow_populated_blocks']}",
        f"Legacy blocks not yet covered: {summary['legacy_blocks_not_yet_covered']}",
        "",
        "Top missing standard blocks:",
    ]
    for row in report["important_legacy_fields_missing_in_shadow"][:20]:
        lines.append(f"- {row['block_id']} ({row['sheet']} {row['range']})")
    if not report["important_legacy_fields_missing_in_shadow"]:
        lines.append("- None from block-level standard comparison.")
    return "\n".join(lines) + "\n"


def run_anf_shadow_workbook_fill(
    *,
    package_path: Path,
    output_dir: Path,
    legacy_workbook_path: Path,
    template_path: Path = DEFAULT_TEMPLATE,
    manifest_path: Path = DEFAULT_MANIFEST,
    binding_map_path: Path = DEFAULT_BINDING_MAP,
    cached_plan_path: Path | None = None,
) -> dict[str, Path]:
    package = _load_json(package_path)
    manifest = _load_json(manifest_path)
    binding_payload = _load_json(binding_map_path)
    bindings = list(binding_payload.get("bindings") or [])
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "workbook": output_dir / "ANF_shadow_model.xlsx",
        "plan_json": output_dir / "ANF_shadow_binding_plan.json",
        "prefill_json": output_dir / "ANF_prefill_coverage_report.json",
        "prefill_txt": output_dir / "ANF_prefill_coverage_report.txt",
        "postfill_json": output_dir / "ANF_shadow_workbook_fill_audit.json",
        "postfill_txt": output_dir / "ANF_shadow_workbook_fill_audit.txt",
        "comparison_json": output_dir / "ANF_shadow_vs_legacy_block_comparison.json",
        "comparison_txt": output_dir / "ANF_shadow_vs_legacy_block_comparison.txt",
    }

    expected_plan_path = cached_plan_path
    if expected_plan_path is None:
        sibling_plan = package_path.with_name("ANF_binding_plan.json")
        expected_plan_path = sibling_plan if sibling_plan.exists() else None
    expected_plan = _load_json(expected_plan_path) if expected_plan_path is not None else None
    plan = reproduce_binding_plan(
        package,
        manifest=manifest,
        binding_payload=binding_payload,
        shell_path=template_path,
        ticker_override="ANF",
        expected_plan=expected_plan,
    )
    write_binding_plan_report(plan, paths["plan_json"])

    prefill = build_prefill_coverage_report(package, bindings, plan=plan)
    _write_json(paths["prefill_json"], prefill)
    _write_text(paths["prefill_txt"], _markdown_prefill(prefill))
    if prefill["status"] != "PASS":
        raise RuntimeError(f"ANF prefill coverage failed: {paths['prefill_json']}")

    candidate_path = _create_candidate_workbook(paths["workbook"])
    try:
        fill_standard_template_from_package(
            package_path,
            output_path=candidate_path,
            ticker_override="ANF",
            template_path=template_path,
            manifest_path=manifest_path,
            binding_map_path=binding_map_path,
            expected_plan=plan,
        )

        strict_validation = validate_shell(
            template_path=candidate_path,
            manifest_path=manifest_path,
            binding_map_path=binding_map_path,
            allow_filled_values=True,
            approved_shell_path=template_path,
            approved_plan_path=paths["plan_json"],
            normalized_package_path=package_path,
        )

        postfill = build_post_fill_audit(
            template_path=template_path,
            output_path=candidate_path,
            manifest=manifest,
            bindings=bindings,
            package=package,
            plan=plan,
            strict_validation=strict_validation,
        )
        postfill["output_path"] = str(paths["workbook"])
        _write_json(paths["postfill_json"], postfill)
        _write_text(paths["postfill_txt"], _markdown_postfill(postfill))
        if strict_validation["status"] != "PASS":
            raise RuntimeError(f"ANF strict post-fill validation failed: {paths['postfill_json']}")

        comparison = build_shadow_vs_legacy_block_comparison(
            template_path=template_path,
            shadow_path=candidate_path,
            legacy_path=legacy_workbook_path,
            prefill_report=prefill,
        )
        _write_json(paths["comparison_json"], comparison)
        _write_text(paths["comparison_txt"], _markdown_comparison(comparison))
        _atomic_promote_workbook(candidate_path, paths["workbook"])
    finally:
        if candidate_path.exists():
            candidate_path.unlink()
    return paths


def main(argv: Sequence[str] | None = None) -> int:
    data_root = _default_data_root()
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--package", type=Path, default=data_root / "outputs" / "stress_tests" / "ANF_new_ticker_engine" / "ANF_normalized_data_package.json")
    parser.add_argument("--output-dir", type=Path, default=data_root / "outputs" / "stress_tests" / "ANF_new_ticker_engine")
    parser.add_argument("--legacy-workbook", type=Path, default=data_root / "outputs" / "Excel stock models" / "ANF_model.xlsx")
    parser.add_argument(
        "--cached-plan",
        type=Path,
        help="Optional serialized plan that must exactly match independent reproduction before fill.",
    )
    args = parser.parse_args(argv)

    paths = run_anf_shadow_workbook_fill(
        package_path=args.package.resolve(),
        output_dir=args.output_dir.resolve(),
        legacy_workbook_path=args.legacy_workbook.resolve(),
        cached_plan_path=args.cached_plan.resolve() if args.cached_plan else None,
    )
    for key, path in paths.items():
        print(f"{key}: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
