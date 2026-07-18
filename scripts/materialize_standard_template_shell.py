"""Materialize the frozen standard stock-model workbook shell.

This is a one-time/repeatable shell authoring helper, not the new-ticker value
filler runtime. It creates an empty visible shell from the manifest and binding
map, using the existing PBI/GPRE/ANF workbooks only as layout contract sources
for broad dimensions/freeze-pane conventions.
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import absolute_coordinate, get_column_letter, quote_sheetname, range_boundaries
from openpyxl.workbook.defined_name import DefinedName

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from pbi_xbrl.standard_template_shell_identity import (
    SHELL_SEMANTIC_CONTRACT_VERSION,
    compute_shell_identity,
    normalize_xlsx_package,
)
from pbi_xbrl.standard_template_formula_contract import (
    FORMULA_CONTRACT_VERSION,
    INVESTMENT_CASE_SCENARIO_OWNED_RANGES,
    INVESTMENT_CASE_SCENARIO_USER_INPUT_RANGES,
    apply_standard_formula_contracts,
    apply_standard_support_formula_contracts,
)
from pbi_xbrl.json_schema_validation import load_json_strict
from pbi_xbrl.standard_template_audit_freshness import write_audit_freshness
from pbi_xbrl.workbook_modules import (
    DEFAULT_MODULE_MANIFEST,
    ResolvedModuleProfile,
    build_profile_binding_payload,
    build_profile_shell_manifest,
    enabled_defined_name_ids,
    enabled_formula_ids,
    load_workbook_module_manifest,
    resolve_module_profile,
    validate_workbook_execution_ownership,
    visible_block_contracts,
)


ROOT = REPO_ROOT
DEFAULT_OUTPUT = ROOT / "templates" / "standard_stock_model_template.xlsx"
DEFAULT_LAB_SOURCE = ROOT / "templates" / "lab" / "ANF_template_lab.xlsx"

SOURCE_TICKERS = ("PBI", "GPRE", "ANF")
SOURCE_SPECIFIC_TERMS = (
    "ANF",
    "Pitney Bowes",
    "Green Plains",
    "Abercrombie",
    "Hollister",
    "Presort",
    "SendTech",
    "45Z",
    "RIN",
    "crush margin",
)
FIXED_DIMENSION_REPLACEMENTS = (
    (re.compile(r"\bAmericas\b", re.I), "[Dimension member 1]"),
    (re.compile(r"\bEMEA\b", re.I), "[Dimension member 2]"),
    (re.compile(r"\bAPAC\b", re.I), "[Dimension member 3]"),
)
FIXED_SECTOR_LABEL_PATTERNS = tuple(
    re.compile(pattern, re.I)
    for pattern in (
        r"\bclosures?\b",
        r"\bopenings?\b",
        r"\bremodels?\b",
        r"\bstores?\s*/\s*buybacks?\b",
        r"\bstores?\s*/\s*real estate\b",
        r"\bfranchise stores?\b",
        r"\bowned stores?\b",
        r"\btariffs?\b",
        r"\bERP\b",
        r"\bfreight tailwind\b",
        r"\bmarketing headwind\b",
        r"\bnet sales growth\b",
        r"\badjusted EPS\b",
        r"\bshare repurchases?\b",
        r"\breal estate activity\b",
    )
)
VALUATION_LABEL_REPLACEMENTS = (
    (re.compile(r"\bAdj\.?\s*EPS\b", re.I), "per-share earnings"),
    (re.compile(r"\badjusted EPS\b", re.I), "per-share earnings"),
    (re.compile(r"\bEPS\b", re.I), "per-share earnings"),
)
SIGNAL_FILL_COLORS = {
    "002F80ED",
    "006FA8DC",
    "009BD3F5",
    "00A63A00",
    "00D55E00",
    "00E69F00",
    "00009E73",
    "0066C2A5",
    "0056B4E9",
    "00CC79A7",
}
GRAY_BLANK_FILLS = {"00DDDDDD", "00D9D9D9", "FFD9D9D9", "FFDDDDDD"}
STATUS_OUTPUT_FILL_COLORS = {"00D9EAF7", "00F2F2F2", "00F4CCCC", "00FFF2CC"}
NEUTRAL_HEADER_FILL_COLORS = SIGNAL_FILL_COLORS
VALUATION_RUNTIME_VALUE_CONSTANT_RANGES = (
    "D194:D216",
    "E236:E240",
    "D247:D250",
    "E253:E256",
    "L248:S250",
)
QA_SHEET_HEADERS = {
    "QA_Log": [
        "issue_id",
        "severity",
        "rule_id",
        "issue_type",
        "section",
        "root_cause",
        "message",
        "suggested_action",
        "occurrence_count",
        "visibility_disposition",
        "promotion_blocking",
        "detail_ref",
    ],
    "Needs_Review": [
        "issue_id",
        "severity",
        "rule_id",
        "section",
        "normalized_path",
        "business_row_key",
        "message",
        "suggested_action",
        "occurrence_count",
        "promotion_blocking",
        "detail_ref",
    ],
    "QA_Checks": [
        "rule_id",
        "status",
        "unique_issue_count",
        "occurrence_count",
        "blocking_count",
        "actionable_count",
        "affected_sections",
        "interpretation",
        "detail_ref",
    ],
}
QA_COLUMN_WIDTHS = {
    "QA_Log": [24, 9, 30, 24, 20, 32, 56, 42, 16, 22, 16, 34],
    "Needs_Review": [24, 9, 30, 20, 32, 24, 56, 42, 16, 16, 34],
    "QA_Checks": [34, 12, 18, 18, 16, 16, 34, 60, 34],
}
VALUATION_GUIDANCE_SIDECAR_HEADERS = {
    "O7": "Guidance",
    "O28": "Metric",
    "Q28": "Stated in",
    "R28": "Applies to",
    "S28": "Guidance",
    "AA28": "Trend / realized",
    "O37": "Operating Drivers",
    "O38": "Driver group",
    "R38": "Driver",
    "U38": "Why it matters",
    "AA38": "Source/type",
    "O48": "Thesis Bridge",
    "O49": "Quick valuation bridge; no market price required.",
    "O50": "Bridge item",
    "U50": "Value",
    "X50": "Notes",
    "O63": "Output",
    "U63": "Value",
    "X63": "Interpretation",
}
VALUATION_STRUCTURAL_HEADERS = {
    "B123": "Principal due ($m)",
    "C123": "Rate type",
    "D123": "Coupon/Spread %",
    "F123": "Maturity",
    "G123": "Conversion price",
    "I123": "Added shares on full conversion (m)",
    "L123": "Concurrent repurchased shares (m)",
    "B138": "Summary",
    "F138": "Score",
    "G138": "Severity",
    "H138": "Result / support",
    "B159": "Δ",
    "C159": "Direction",
    "D159": "As-of",
    "B169": "Status",
    "C169": "Evidence",
    "I169": "As-of",
}
VALUATION_BLUE_SECTION_HEADERS = {
    "O7": "Guidance",
    "O37": "Operating Drivers",
    "O48": "Thesis Bridge",
    "A122": "Debt Detail (latest)",
    "A137": "Hidden value flags",
    "N137": "Hidden Value Panel",
    "A145": "Operating signals",
    "A151": "Capital return",
    "A158": "Trend/Δ (last 4Q)",
    "A168": "Red/Green Flags",
    "B192": "Valuation",
}
STANDARD_RED_GREEN_FLAG_LABELS = {
    170: "Red: Revenue up but CFO down (YoY)",
    171: "Red: Earnings quality CFO/NI (TTM)",
    172: "Red: AR growing faster than revenue (YoY)",
    173: "Red: Inventory build without revenue growth",
    174: "Red: Debt growing faster than revenue (YoY)",
    175: "Red: Leverage rising (YoY Δ)",
    176: "Red: Interest coverage low (cash)",
    177: "Red: FCF negative while EBITDA positive (TTM)",
    178: "Watch: Buybacks exceeded FCF",
    179: "Red: Goodwill heavy",
    180: "Red: Share dilution (YoY)",
    181: "Red: Pension obligations pressure",
    183: "Green: Operating margin trend QoQ",
    184: "Green: FCF TTM growth (YoY)",
    185: "Green: Net debt decreasing (YoY)",
    186: "Green: Interest coverage improving (YoY)",
    187: "Green: Shares outstanding decreasing (YoY)",
    188: "Green: Liquidity improving (YoY)",
}
OPERATING_DRIVER_SHEET_HEADERS = {
    "A12": "Topic",
    "B12": "Current read",
    "H12": "Source / use",
    "A19": "Horizon",
    "B19": "Stated in",
    "C19": "Commentary",
}
SHEET_REF_RE = re.compile(r"'([^']+)'!|(?<![A-Za-z0-9_])([A-Za-z_][A-Za-z0-9_ ]{0,60})!")
COMPANY_SPECIFIC_DEFINED_NAMES = {"ThesisBaseAdjEBITDA_FY"}
STALE_UNREFERENCED_DEFINED_NAMES = {
    "FCF_Yield",
    "summary_net_debt",
    "summary_quarterly_revenue",
    "summary_revenue_model",
}
NEUTRAL_STATIC_LABELS = {
    "SUMMARY": {
        "A45": "Total liquidity",
        "C45": "$m",
    },
    "Valuation": {
        "A36": "Net income",
        "A37": "Net margin %",
        "A38": "Net income YoY %",
        "A39": "Net income (TTM)",
        "A40": "Net margin (TTM)",
    },
    "BS_Segments": {
        "A2": "Quarter labels are fiscal periods. Values are scaled to $m unless otherwise stated.",
        "A3": "QA checks populate after normalized data is planned.",
    },
    "Operating_Drivers": {
        "A13": "[Current actual period]",
        "A14": "[Current guidance period]",
        "A21": "[Operating evidence period slot 1]",
        "A22": "[Operating evidence period slot 2]",
        "A23": "[Operating evidence period slot 3]",
        "A25": "[Operating evidence period slot 4]",
        "A26": "[Operating evidence period slot 5]",
        "A27": "[Operating evidence period slot 6]",
        "A28": "[Operating evidence period slot 7]",
        "A29": "[Operating evidence period slot 8]",
        "A30": "[Operating evidence period slot 9]",
        "A32": "[Operating evidence period slot 10]",
        "A33": "[Operating evidence period slot 11]",
        "A34": "[Operating evidence period slot 12]",
        "A35": "[Operating evidence period slot 13]",
        "A36": "[Operating evidence period slot 14]",
        "A37": "[Operating evidence period slot 15]",
        "A38": "[Operating evidence period slot 16]",
        "A41": "Segment support - latest 12 periods",
        "A48": "[Operating scope slot 48]",
        "A52": "Actuals - latest 12 periods",
        "A64": "Dimension view",
        "A70": "[Optional sector metric slot 70]",
        "A71": "[Optional sector metric slot 71]",
        "A73": "[Optional sector metric slot 73]",
        "A74": "[Optional sector metric slot 74]",
        "A76": "[Optional sector metric slot 76]",
        "A92": "Outlook bridge",
        "A110": "[Optional sector operating driver slot 110]",
        "A111": "[Optional sector block slot 111]",
        "A112": "[Optional sector metric slot 112]",
    },
    "{ticker}_Investment_Case": {
        "A39": "Dimension view 1 - summed only when selected",
        "A42": "Dimension view 2 - summed only when selected",
        "A73": "[Key debate slot 1]",
        "A79": "[Key debate slot 2]",
        "A92": "[Quality of earnings item 1]",
        "A93": "[Quality of earnings item 2]",
        "A96": "[Quality of earnings item 5]",
        "A100": "[Scenario condition slot 1]",
        "A101": "[Scenario condition slot 2]",
        "A102": "[Scenario condition slot 3]",
        "A103": "[Scenario condition slot 4]",
        "A104": "[Scenario condition slot 5]",
        "A108": "[Margin baseline slot]",
        "A109": "[Margin guidance slot]",
        "A110": "[Margin delta slot]",
        "A114": "[Margin / cost driver slot 1]",
        "A115": "[Margin / realization driver slot]",
        "A116": "[Margin / cost driver slot 2]",
        "A126": "Guided EPS",
        "A132": "Buybacks",
        "A138": "Guidance to implied earnings",
        "A140": "Revenue baseline",
        "A141": "Revenue growth guide",
        "A142": "Implied revenue",
        "A153": "[Operating margin sensitivity slot]",
        "A154": "[Gross margin sensitivity slot]",
        "A155": "[Revenue growth sensitivity slot]",
        "A156": "[Capital return sensitivity slot]",
        "A157": "[EPS sensitivity slot]",
        "A184": "[Comparison period slot]",
        "A193": "Business Health",
        "A195": "Sales",
        "A196": "Sales growth",
        "A197": "Current-period sales growth",
        "A198": "[Current-period operating metric slot]",
        "A201": "Inventory / Working-Capital Risk",
        "A211": "Asset Productivity / Capacity Returns",
        "A215": "[Asset scope slot]",
        "A223": "[Asset growth slot]",
        "A224": "Revenue growth vs asset growth",
        "A225": "[Channel mix slot]",
    },
    "Promise_Progress_UI": {
        "A11": "Guidance progression - period block 1",
        "A22": "Guidance progression - period block 2",
        "A28": "Guidance progression - period block 3",
        "A33": "Guidance progression - period block 4",
        "A37": "Current open guidance",
        "A59": "Guidance timeline block 1",
        "A69": "Guidance timeline block 2",
        "A76": "Guidance timeline block 3",
        "A84": "Guidance timeline block 4",
        "A90": "Guidance timeline block 5",
        "A97": "Guidance timeline block 6",
    },
}
VISIBLE_SOURCE_TEXT_PATTERNS = tuple(
    re.compile(pattern, re.I)
    for pattern in (
        r"\bsource-backed\b",
        r"\bearnings release\b",
        r"\bguidance profile\b",
        r"\bmodel_metric\b",
        r"\bpre-release update\b",
        r"StockModelData[\\/]tickers[\\/]",
        r"\b20\d{2}(?:-Q[1-4]|-\d{2}-\d{2})?\b",
        r"\$\s*-?\d",
        r"\b\d+(?:\.\d+)?\s*(?:%|bps|million|billion)\b",
        r"\brevolver(?:_|\s+)capacity(?:_|\s+)change",
        r"\bbrand_family_momentum\b",
        r"\bAUR\b",
    )
)
QUARTER_NOTES_STATIC_TEXT = {
    "quarter read",
    "model read",
    "what changed",
    "watch next",
    "key caveat",
    "key developments",
    "theme",
    "what happened",
    "why it matters",
    "model / valuation implication",
    "source / confidence",
    "guidance / promise interpretation",
    "promise / guidance item",
    "read",
    "actual / progress interpretation",
    "status / caveat",
    "source",
    "model mapping / double-count guardrails",
    "driver",
    "model treatment",
    "double-count guardrail",
    "linked sheet / metric",
}
QUARTER_NOTES_BLOCK_TITLE_RE = re.compile(r"^20\d{2}-Q[1-4]\s+-\s+Quarter Notes$", re.I)
NEUTRAL_SLOT_RE = re.compile(r"^\[(?:Quarter note theme|Dimension member) slot?\s*\d+\]$", re.I)


@dataclass(frozen=True)
class SourceSheetContract:
    freeze_panes: str | None
    column_widths: dict[str, float]
    row_heights: dict[int, float]


def _load_json(path: Path) -> dict[str, Any]:
    payload = load_json_strict(path)
    if not isinstance(payload, dict):
        raise ValueError(f"JSON contract must be an object: {path}")
    return payload


def _default_data_root() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = parent / "StockModelData"
        if candidate.exists():
            return candidate
    return ROOT.parent / "StockModelData"


DEFAULT_DATA_ROOT = _default_data_root()


def _source_workbook_paths(data_root: Path) -> dict[str, Path]:
    output_dir = data_root / "outputs" / "Excel stock models"
    return {ticker: output_dir / f"{ticker}_model.xlsx" for ticker in SOURCE_TICKERS}


def _source_sheet_name(sheet_name: str, ticker: str) -> str:
    if sheet_name == "{ticker}_Investment_Case":
        return f"{ticker}_Investment_Case"
    return sheet_name


def _load_source_contracts(data_root: Path, manifest: dict[str, Any]) -> dict[str, SourceSheetContract]:
    paths = _source_workbook_paths(data_root)
    workbooks: dict[str, Any] = {}
    for ticker, path in paths.items():
        if path.exists():
            workbooks[ticker] = load_workbook(path, read_only=False, data_only=False)

    contracts: dict[str, SourceSheetContract] = {}
    for sheet_def in manifest["sheets"]:
        sheet_name = str(sheet_def["sheet"])
        chosen = None
        for ticker in SOURCE_TICKERS:
            wb = workbooks.get(ticker)
            candidate = _source_sheet_name(sheet_name, ticker)
            if wb is not None and candidate in wb.sheetnames:
                chosen = wb[candidate]
                break
        if chosen is None:
            contracts[sheet_name] = SourceSheetContract(
                freeze_panes=None,
                column_widths={},
                row_heights={},
            )
            continue
        contracts[sheet_name] = SourceSheetContract(
            freeze_panes=str(chosen.freeze_panes) if chosen.freeze_panes else None,
            column_widths={
                col: float(dim.width)
                for col, dim in chosen.column_dimensions.items()
                if dim.width is not None and 1.0 <= float(dim.width) <= 80.0
            },
            row_heights={
                int(row): float(dim.height)
                for row, dim in chosen.row_dimensions.items()
                if dim.height is not None and 1 <= int(row) <= 1000
            },
        )
    for wb in workbooks.values():
        wb.close()
    return contracts


def _max_manifest_bounds(sheet_def: dict[str, Any]) -> tuple[int, int]:
    max_col = 1
    max_row = 1
    for zone_type in ("writable_zones", "non_writable_zones"):
        for zone in sheet_def[zone_type]:
            min_col, min_row, max_col_in, max_row_in = range_boundaries(zone["target"])
            max_col = max(max_col, max_col_in, min_col)
            max_row = max(max_row, max_row_in, min_row)
    return max_col, max_row


def _configure_investment_case_ownership_zones(manifest: dict[str, Any]) -> None:
    sheet = next(
        (row for row in manifest.get("sheets") or [] if row.get("sheet") == "{ticker}_Investment_Case"),
        None,
    )
    if not isinstance(sheet, dict):
        return
    retained = [
        zone
        for zone in sheet.get("writable_zones") or []
        if str(zone.get("zone_id") or "")
        not in {"ic_key_debate_values", "ic_manual_input_values", "ic_scenario_bridge_values"}
    ]
    snapshot = next((zone for zone in retained if zone.get("zone_id") == "ic_snapshot_values"), None)
    if snapshot is not None:
        snapshot["target"] = "B5:B11"
    scenario_inputs = [
        {
            "zone_id": f"ic_scenario_user_input_{index}",
            "target": target,
            "anchor_label": "Typed Scenario Inputs",
            "value_shapes": ["scalar", "table_rows"],
        }
        for index, target in enumerate(INVESTMENT_CASE_SCENARIO_USER_INPUT_RANGES, start=1)
    ]
    sheet["writable_zones"] = [*retained, *scenario_inputs]
    non_writable = [
        zone
        for zone in sheet.get("non_writable_zones") or []
        if str(zone.get("zone_id") or "") != "ic_static_label_column"
    ]
    non_writable.extend(
        {
            "zone_id": f"ic_static_label_column_{index}",
            "target": target,
            "reason": "Static thesis/scenario labels next to exact writable inputs.",
        }
        for index, target in enumerate(("A5:A160", "A164:A171", "A175:A177", "A181:A184"), start=1)
    )
    sheet["non_writable_zones"] = non_writable


def _ranges_for(sheet_def: dict[str, Any], zone_type: str) -> list[str]:
    return [str(zone["target"]) for zone in sheet_def.get(zone_type, [])]


def _sheet_title(sheet_name: str) -> str:
    return {
        "SUMMARY": "SUMMARY",
        "Valuation": "Valuation",
        "BS_Segments": "Balance Sheet & Segments",
        "Operating_Drivers": "Operating Drivers",
        "{ticker}_Investment_Case": "",
        "Quarter_Notes_UI": "Quarter Notes",
        "Promise_Progress_UI": "Promise Progress",
        "QA_Log": "QA Log",
        "Needs_Review": "Needs Review",
        "QA_Checks": "QA Checks",
    }.get(sheet_name, sheet_name)


def _title_cell(sheet_name: str) -> str:
    return {
        "Valuation": "A3",
        "Operating_Drivers": "A2",
    }.get(sheet_name, "A1")


def _fallback_freeze(sheet_name: str) -> str:
    return {
        "SUMMARY": "A4",
        "Valuation": "B20",
        "BS_Segments": "B8",
        "Operating_Drivers": "B7",
        "{ticker}_Investment_Case": "B5",
        "Quarter_Notes_UI": "B8",
        "Promise_Progress_UI": "B8",
        "QA_Log": "A2",
        "Needs_Review": "A2",
        "QA_Checks": "A2",
    }.get(sheet_name, "A2")


def _anchor_row_from_zone(sheet_def: dict[str, Any], zone_id: str) -> int:
    for zone in sheet_def["writable_zones"]:
        if zone["zone_id"] == zone_id:
            _min_col, min_row, _max_col, _max_row = range_boundaries(zone["target"])
            return int(min_row)
    return 1


def _binding_anchor_rows(bindings: Iterable[dict[str, Any]]) -> dict[tuple[str, int], str]:
    labels: dict[tuple[str, int], str] = {}
    for entry in bindings:
        sheet = str(entry["sheet"])
        target = str(entry["target"])
        _min_col, min_row, _max_col, _max_row = range_boundaries(target)
        if str(entry.get("source_policy")) == "validation-output":
            continue
        label = str(entry.get("anchor_label") or entry.get("section") or "").strip()
        if label:
            labels[(sheet, int(min_row))] = label
    return labels


def _apply_zone_style(ws: Any, ranges: Iterable[str], fill: PatternFill, border: Border) -> None:
    for range_ref in ranges:
        for row in ws[range_ref]:
            for cell in row:
                cell.fill = fill
                cell.border = border


def _merge_title(ws: Any, sheet_name: str, max_col: int) -> None:
    if sheet_name in {"QA_Log", "Needs_Review", "QA_Checks"}:
        return
    title_ref = _title_cell(sheet_name)
    row = int("".join(ch for ch in title_ref if ch.isdigit()))
    end_col = min(max_col, 15)
    if end_col > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)


def _apply_dimensions(ws: Any, sheet_name: str, contract: SourceSheetContract, max_col: int, max_row: int) -> None:
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        width = contract.column_widths.get(letter)
        if width is None:
            if col_idx == 1:
                width = 28.0
            elif sheet_name in {"Quarter_Notes_UI", "Promise_Progress_UI"}:
                width = 18.0
            elif sheet_name in {"QA_Log", "Needs_Review", "QA_Checks"}:
                width = 20.0
            else:
                width = 14.0
        ws.column_dimensions[letter].width = width
    for row_idx in range(1, min(max_row, 260) + 1):
        height = contract.row_heights.get(row_idx)
        if height is None:
            height = 24.0 if row_idx in {1, 2, 3} else 18.0
        ws.row_dimensions[row_idx].height = height
    ws.freeze_panes = contract.freeze_panes or _fallback_freeze(sheet_name)


def _write_static_structure(
    wb: Workbook,
    ws: Any,
    sheet_def: dict[str, Any],
    bindings: list[dict[str, Any]],
    contract: SourceSheetContract,
) -> None:
    sheet_name = str(sheet_def["sheet"])
    max_col, max_row = _max_manifest_bounds(sheet_def)

    title_fill = PatternFill("solid", fgColor="4472C4")
    section_fill = PatternFill("solid", fgColor="DDEBF7")
    label_fill = PatternFill("solid", fgColor="F2F6FA")
    writable_fill = PatternFill("solid", fgColor="FFFFFF")
    border = Border(
        left=Side(style="thin", color="D9E2EA"),
        right=Side(style="thin", color="D9E2EA"),
        top=Side(style="thin", color="D9E2EA"),
        bottom=Side(style="thin", color="D9E2EA"),
    )

    ws.sheet_view.showGridLines = False
    _apply_dimensions(ws, sheet_name, contract, max_col, max_row)
    _apply_zone_style(ws, _ranges_for(sheet_def, "writable_zones"), writable_fill, border)
    _apply_zone_style(ws, _ranges_for(sheet_def, "non_writable_zones"), label_fill, border)

    _merge_title(ws, sheet_name, max_col)
    title_cell = ws[_title_cell(sheet_name)]
    title_cell.value = _sheet_title(sheet_name)
    title_cell.fill = title_fill
    title_cell.font = Font(bold=True, color="FFFFFF", size=15)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    binding_rows = _binding_anchor_rows(bindings)
    for (binding_sheet, row_idx), label in binding_rows.items():
        if binding_sheet != sheet_name:
            continue
        cell = ws.cell(row_idx, 1)
        if not cell.value:
            cell.value = label
        cell.fill = section_fill
        cell.font = Font(bold=True, color="1F2933")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for label in sheet_def.get("formulas_static_labels", []):
        label_text = str(label or "").strip()
        if not label_text:
            continue
        if any(str(cell.value or "").strip() == label_text for row in ws.iter_rows() for cell in row):
            continue
        # Place extra static labels in the first label column row that is still blank.
        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row_idx, 1)
            if not cell.value:
                cell.value = label_text
                cell.fill = section_fill
                cell.font = Font(bold=True, color="1F2933")
                break

    if sheet_name in {"QA_Log", "Needs_Review", "QA_Checks"}:
        ws["A1"] = _sheet_title(sheet_name)
        for col_idx, header in enumerate(QA_HEADERS, start=2):
            ws.cell(1, col_idx, header)
        for cell in ws[1]:
            cell.fill = title_fill if cell.column == 1 else section_fill
            cell.font = Font(bold=True, color="FFFFFF" if cell.column == 1 else "1F2933")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    if sheet_name == "Valuation":
        ws["A1"] = "Scale"
        ws["B1"] = "$m"
        ws["A2"] = "Values scaled to $m unless %"

    for anchor in wb._standard_template_required_anchors:  # type: ignore[attr-defined]
        if anchor["sheet"] != sheet_name:
            continue
        row_idx = 1 if sheet_name in {"QA_Log", "Needs_Review", "QA_Checks"} else _anchor_row_from_zone(sheet_def, anchor["zone_id"])
        cell = ws.cell(row_idx, 1)
        if not cell.value:
            cell.value = anchor.get("anchor_label") or anchor["anchor_id"]
        coord = absolute_coordinate(cell.coordinate)
        attr_text = f"{quote_sheetname(sheet_name)}!{coord}"
        wb.defined_names.add(DefinedName(str(anchor["anchor_id"]), attr_text=attr_text))


def _rich_template_source(data_root: Path) -> Path | None:
    if DEFAULT_LAB_SOURCE.exists():
        return DEFAULT_LAB_SOURCE
    source = data_root / "outputs" / "Excel stock models" / "ANF_model.xlsx"
    return source if source.exists() else None


def _rename_investment_case_sheet(wb: Workbook) -> None:
    if "ANF_Investment_Case" in wb.sheetnames and "{ticker}_Investment_Case" not in wb.sheetnames:
        wb["ANF_Investment_Case"].title = "{ticker}_Investment_Case"
    if "ANF_Investment_Case_Data" in wb.sheetnames and "{ticker}_Investment_Case_Data" not in wb.sheetnames:
        wb["ANF_Investment_Case_Data"].title = "{ticker}_Investment_Case_Data"


def _hide_nonstandard_sheets(wb: Workbook, manifest: dict[str, Any]) -> None:
    states = {
        str(sheet_name): str(state)
        for sheet_name, state in (manifest.get("module_profile", {}).get("sheet_states", {}) or {}).items()
    }
    standard_visible = {str(sheet_name) for sheet_name in manifest["visible_sheet_order"]}
    for ws in wb.worksheets:
        ws.sheet_state = states.get(ws.title, "visible" if ws.title in standard_visible else "hidden")


def _clear_range_values(ws: Any, range_ref: str) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _clear_writable_zones(wb: Workbook, manifest: dict[str, Any]) -> None:
    for sheet_def in manifest["sheets"]:
        sheet_name = str(sheet_def["sheet"])
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for zone in sheet_def["writable_zones"]:
            _clear_range_values(ws, str(zone["target"]))


def _clear_visible_source_values_and_notes(wb: Workbook, manifest: dict[str, Any]) -> None:
    """Remove source-backed payload while retaining labels, formulas, and styles."""

    for sheet_name in manifest["visible_sheet_order"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                value = cell.value
                if isinstance(value, (int, float, date, datetime)) and not isinstance(value, bool):
                    cell.value = None
                    continue
                if isinstance(value, str) and not value.startswith("=") and any(pattern.search(value) for pattern in VISIBLE_SOURCE_TEXT_PATTERNS):
                    cell.value = None


def _clear_investment_case_scenario_surfaces(wb: Workbook) -> None:
    sheet_name = "{ticker}_Investment_Case"
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    for range_ref in INVESTMENT_CASE_SCENARIO_OWNED_RANGES:
        for row in ws[range_ref]:
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.comment = None
                cell.hyperlink = None


def _neutralize_quarter_notes_history(wb: Workbook) -> None:
    """Keep reusable history-block headers while removing inherited evidence."""

    sheet_name = "Quarter_Notes_UI"
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    block_index = 0
    for row in ws.iter_rows(min_row=16):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            value = cell.value
            if value in (None, "") or (isinstance(value, str) and value.startswith("=")):
                continue
            text = str(value).strip()
            if cell.column == 1 and QUARTER_NOTES_BLOCK_TITLE_RE.fullmatch(text):
                block_index += 1
                cell.value = f"[Historical quarter block {block_index}]"
                continue
            if text.casefold() in QUARTER_NOTES_STATIC_TEXT:
                continue
            if cell.column == 1 and NEUTRAL_SLOT_RE.fullmatch(text):
                continue
            if cell.column == 1:
                cell.value = f"[Quarter note theme slot {cell.row}]"
                continue
            cell.value = None


def _clear_source_specific_visible_text(wb: Workbook, manifest: dict[str, Any]) -> None:
    term_re = re.compile("|".join(r"\b" + re.escape(term) + r"\b" for term in SOURCE_SPECIFIC_TERMS), re.I)
    for sheet_name in manifest["visible_sheet_order"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell) or not isinstance(cell.value, str):
                    continue
                if cell.value.startswith("="):
                    continue
                if term_re.search(cell.value):
                    cell.value = None


def _cell_row(coord: str) -> int:
    digits = "".join(ch for ch in coord if ch.isdigit())
    return int(digits or "0")


def _generic_slot_label(sheet_name: str, coord: str, original: str) -> str:
    row = _cell_row(coord)
    if sheet_name == "Valuation":
        text = original
        for pattern, replacement in VALUATION_LABEL_REPLACEMENTS:
            text = pattern.sub(replacement, text)
        return text
    if sheet_name == "BS_Segments":
        return f"[Dimension member slot {row}]"
    if sheet_name == "Operating_Drivers":
        return f"[Operating driver slot {row}]"
    if sheet_name == "{ticker}_Investment_Case":
        return f"[Scenario driver slot {row}]"
    if sheet_name == "Quarter_Notes_UI":
        return f"[Quarter note theme slot {row}]"
    if sheet_name == "Promise_Progress_UI":
        return f"[Guidance metric slot {row}]"
    return f"[Template slot {row}]"


def _genericize_visible_text_value(sheet_name: str, coord: str, value: str) -> str:
    text = value
    if sheet_name == "Valuation":
        for pattern, replacement in VALUATION_LABEL_REPLACEMENTS:
            text = pattern.sub(replacement, text)
    had_dimension_member = False
    for pattern, replacement in FIXED_DIMENSION_REPLACEMENTS:
        if pattern.search(text):
            had_dimension_member = True
        text = pattern.sub(replacement, text)

    had_sector_label = any(pattern.search(text) for pattern in FIXED_SECTOR_LABEL_PATTERNS) or any(
        pattern.search(value) for pattern in FIXED_SECTOR_LABEL_PATTERNS
    )
    if had_sector_label:
        return _generic_slot_label(sheet_name, coord, text)
    if had_dimension_member:
        text = re.sub(r"\s*\(geography\s*/\s*stores\)", "", text, flags=re.I)
        text = re.sub(r"\bstores?\b", "metric", text, flags=re.I)
        return text
    return text


def _genericize_sector_specific_visible_text(wb: Workbook, manifest: dict[str, Any]) -> None:
    for sheet_name in manifest["visible_sheet_order"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell) or not isinstance(cell.value, str):
                    continue
                if cell.value.startswith("="):
                    continue
                replacement = _genericize_visible_text_value(sheet_name, cell.coordinate, cell.value)
                if replacement != cell.value:
                    cell.value = replacement


def _clear_valuation_numeric_constants(wb: Workbook) -> None:
    if "Valuation" not in wb.sheetnames:
        return
    ws = wb["Valuation"]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, (int, float, date, datetime)):
                cell.value = None


def _clear_valuation_runtime_value_constants(wb: Workbook) -> None:
    if "Valuation" not in wb.sheetnames:
        return
    ws = wb["Valuation"]
    for range_ref in VALUATION_RUNTIME_VALUE_CONSTANT_RANGES:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value in (None, ""):
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                cell.value = None


def _neutralize_writable_data_like_fills(wb: Workbook, manifest: dict[str, Any]) -> None:
    neutral_fill = PatternFill(fill_type=None)
    for sheet_def in manifest["sheets"]:
        sheet_name = str(sheet_def["sheet"])
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for zone in sheet_def["writable_zones"]:
            min_col, min_row, max_col, max_row = range_boundaries(str(zone["target"]))
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    if isinstance(cell, MergedCell) or cell.value not in (None, ""):
                        continue
                    fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == "rgb" else ""
                    if fill in SIGNAL_FILL_COLORS or fill in GRAY_BLANK_FILLS:
                        cell.fill = neutral_fill


def _neutralize_visible_blank_gray_fills(wb: Workbook) -> None:
    neutral_fill = PatternFill(fill_type=None)
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell) or cell.value not in (None, ""):
                    continue
                fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == "rgb" else ""
                if fill in GRAY_BLANK_FILLS:
                    cell.fill = neutral_fill


def _neutralize_valuation_signal_fills(wb: Workbook) -> None:
    if "Valuation" not in wb.sheetnames:
        return
    ws = wb["Valuation"]
    neutral_fill = PatternFill(fill_type=None)
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.coordinate in VALUATION_BLUE_SECTION_HEADERS:
                continue
            fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == "rgb" else ""
            if fill in SIGNAL_FILL_COLORS:
                cell.fill = neutral_fill


def _neutralize_red_green_flags(wb: Workbook) -> None:
    if "Valuation" not in wb.sheetnames:
        return
    ws = wb["Valuation"]
    neutral_fill = PatternFill(fill_type=None)
    ws["A168"] = VALUATION_BLUE_SECTION_HEADERS["A168"]
    ws["A169"] = "Flag"
    ws["B169"] = "Status"
    ws["C169"] = "Evidence"
    ws["I169"] = "As-of"
    for row_idx in range(170, 189):
        label_cell = ws.cell(row_idx, 1)
        if not isinstance(label_cell, MergedCell):
            label_cell.value = STANDARD_RED_GREEN_FLAG_LABELS.get(row_idx)
        for col_idx in range(2, 10):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == "rgb" else ""
            if fill in SIGNAL_FILL_COLORS or fill in STATUS_OUTPUT_FILL_COLORS or fill in GRAY_BLANK_FILLS:
                cell.fill = neutral_fill


def _neutralize_blank_valuation_value_fills(wb: Workbook) -> None:
    if "Valuation" not in wb.sheetnames:
        return
    ws = wb["Valuation"]
    neutral_fill = PatternFill(fill_type=None)
    for range_ref in ("U51:U62",):
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if isinstance(cell, MergedCell) or cell.value not in (None, ""):
                    continue
                fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == "rgb" else ""
                if fill not in ("", "00000000", "00FFFFFF", "FFFFFFFF"):
                    cell.fill = neutral_fill


def _ensure_promise_progress_structure(wb: Workbook) -> None:
    if "Promise_Progress_UI" not in wb.sheetnames:
        return
    ws = wb["Promise_Progress_UI"]
    annual_headers = [
        "Metric",
        "Initial guide",
        "Q1 update",
        "Q2 update",
        "Q3 update",
        "Q4 update",
        "Actual",
        "Status",
        "Notes/source",
    ]
    open_guidance_headers = ["Metric", "Current guide", "Horizon", "Status", "Notes/source"]
    revision_headers = [
        "Metric",
        "Previous guide",
        "New/current guide",
        "Change type",
        "Actual",
        "Progress / run-rate",
        "Status",
        "Horizon",
        "Stated in",
        "Source date",
        "Source / note",
    ]

    for row_idx in (12, 23, 29, 34):
        for col_idx, header in enumerate(annual_headers, start=1):
            ws.cell(row_idx, col_idx, header)
    for col_idx, header in enumerate(open_guidance_headers, start=1):
        ws.cell(38, col_idx, header)
    for row_idx in (60, 70, 77, 85, 91, 98):
        for col_idx, header in enumerate(revision_headers, start=1):
            ws.cell(row_idx, col_idx, header)

    for slot_idx, row_idx in enumerate((5, 6, 7, 8, 9), start=1):
        ws.cell(row_idx, 1, f"[Credibility category slot {slot_idx}]")
    guidance_rows = [
        14,
        15,
        16,
        17,
        18,
        24,
        25,
        26,
        27,
        30,
        31,
        32,
        35,
        39,
        40,
        41,
        42,
        43,
        44,
        45,
        46,
        47,
        48,
        49,
        50,
        51,
        52,
        53,
        54,
        55,
        56,
        57,
        58,
        59,
        61,
        62,
        63,
        64,
        65,
        66,
        67,
        68,
        69,
        71,
        72,
        73,
        74,
        75,
        76,
        78,
        79,
        80,
        81,
        82,
        83,
        84,
        86,
        87,
        88,
        89,
        90,
        92,
        93,
        94,
        95,
        96,
        97,
        99,
        100,
        101,
        102,
        103,
        104,
        105,
        106,
        107,
        108,
        109,
        110,
        111,
        112,
        113,
        114,
        115,
    ]
    for row_idx in guidance_rows:
        cell = ws.cell(row_idx, 1)
        if isinstance(cell, MergedCell):
            continue
        if cell.value in (None, ""):
            continue
        text = str(cell.value)
        if text == "Metric" or text.startswith("["):
            continue
        cell.value = f"[Guidance metric slot {row_idx}]"


def _configure_narrative_text_layout(wb: Workbook) -> None:
    if "Operating_Drivers" in wb.sheetnames:
        ws = wb["Operating_Drivers"]
        for row_idx in (6, 7, 8, 9, 14, 15):
            alignment = copy(ws[f"B{row_idx}"].alignment)
            alignment.wrap_text = True
            alignment.vertical = "center"
            ws[f"B{row_idx}"].alignment = alignment
            ws.row_dimensions[row_idx].height = 42.0

    if "Promise_Progress_UI" in wb.sheetnames:
        ws = wb["Promise_Progress_UI"]
        for row_idx, columns, height in ((19, "BCDE", 60.0), (67, "BC", 42.0)):
            for column in columns:
                alignment = copy(ws[f"{column}{row_idx}"].alignment)
                alignment.wrap_text = True
                alignment.vertical = "center"
                ws[f"{column}{row_idx}"].alignment = alignment
            ws.row_dimensions[row_idx].height = height


def _ensure_merged_range(ws: Any, range_ref: str) -> None:
    if range_ref not in {str(merged_range) for merged_range in ws.merged_cells.ranges}:
        ws.merge_cells(range_ref)


def _style_cells(
    ws: Any,
    range_ref: str,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    alignment: Alignment | None = None,
) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if fill is not None:
                cell.fill = copy(fill)
            if font is not None:
                cell.font = copy(font)
            if alignment is not None:
                cell.alignment = copy(alignment)


def _ensure_valuation_guidance_sidecar_headers(wb: Workbook) -> None:
    if "Valuation" not in wb.sheetnames:
        return
    ws = wb["Valuation"]
    section_fill = PatternFill("solid", fgColor="6FA8DC")
    column_header_fill = PatternFill("solid", fgColor="EAF3FB")
    section_font = Font(bold=True, color="FFFFFF", size=12)
    header_font = Font(bold=True, color="000000", size=12)
    title_font = Font(bold=True, color="FFFFFF", size=18)
    left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_center = Alignment(horizontal="center", vertical="center")

    for range_ref in ("D123:E123", "H138:I138"):
        _ensure_merged_range(ws, range_ref)

    for coord, value in {**VALUATION_GUIDANCE_SIDECAR_HEADERS, **VALUATION_STRUCTURAL_HEADERS}.items():
        cell = ws[coord]
        if not isinstance(cell, MergedCell):
            cell.value = value
            cell.fill = column_header_fill
            cell.font = header_font
            cell.alignment = left_center

    for coord, value in VALUATION_BLUE_SECTION_HEADERS.items():
        cell = ws[coord]
        if not isinstance(cell, MergedCell):
            cell.value = value
            cell.fill = section_fill
            cell.font = section_font
            cell.alignment = left_center

    _style_cells(ws, "A122:N122", fill=section_fill)
    _style_cells(ws, "B123:N123", fill=column_header_fill, font=header_font, alignment=left_center)
    _style_cells(ws, "B138:I138", fill=column_header_fill, font=header_font, alignment=left_center)
    _style_cells(ws, "A145:M145", fill=section_fill)
    _style_cells(ws, "A151:M151", fill=section_fill)
    _style_cells(ws, "A158:D158", fill=section_fill)
    _style_cells(ws, "A169:I169", fill=column_header_fill, font=header_font, alignment=center_center)
    _style_cells(ws, "B192:S192", fill=section_fill, font=title_font)

    for coord in ("A122", "A145", "A151", "A158", "A168"):
        ws[coord].font = section_font
    ws["B192"].value = "Valuation"
    ws["B192"].font = title_font


def _ensure_operating_driver_sheet_headers(wb: Workbook) -> None:
    if "Operating_Drivers" not in wb.sheetnames:
        return
    ws = wb["Operating_Drivers"]
    for coord, value in OPERATING_DRIVER_SHEET_HEADERS.items():
        cell = ws[coord]
        if not isinstance(cell, MergedCell):
            cell.value = value


def _move_operating_drivers_title_to_row1(wb: Workbook) -> None:
    if "Operating_Drivers" not in wb.sheetnames:
        return
    ws = wb["Operating_Drivers"]
    title_fill = PatternFill("solid", fgColor="6FA8DC")
    title_font = Font(bold=True, color="FFFFFF", size=15)
    title_alignment = Alignment(horizontal="center", vertical="center")

    if "A2:N2" in {str(merged_range) for merged_range in ws.merged_cells.ranges}:
        ws.unmerge_cells("A2:N2")
    _ensure_merged_range(ws, "A1:N1")

    ws["A1"] = "Operating Drivers"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = title_alignment
    ws.row_dimensions[1].height = ws.row_dimensions[2].height or 24

    for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=14):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
    ws.freeze_panes = "A2"


def _neutralize_remaining_visible_row_labels(wb: Workbook) -> None:
    replacements: dict[str, list[int]] = {
        "Operating_Drivers": [6, 31, 56, 87, 88, 89, 102],
        "Quarter_Notes_UI": [109, 140, 228],
    }
    for sheet_name, rows in replacements.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_idx in rows:
            cell = ws.cell(row_idx, 1)
            if isinstance(cell, MergedCell) or cell.value in (None, ""):
                continue
            cell.value = _generic_slot_label(sheet_name, cell.coordinate, str(cell.value))


def _ensure_sheet_titles(wb: Workbook, manifest: dict[str, Any]) -> None:
    for sheet_name in manifest["visible_sheet_order"]:
        if sheet_name not in wb.sheetnames:
            continue
        cell = wb[sheet_name][_title_cell(str(sheet_name))]
        if not isinstance(cell, MergedCell):
            cell.value = _sheet_title(str(sheet_name))


def _ensure_binding_anchor_labels(wb: Workbook, manifest: dict[str, Any], bindings: list[dict[str, Any]]) -> None:
    current_binding_ids = {str(entry["binding_id"]) for entry in bindings}
    current_anchor_ids = {str(anchor["anchor_id"]) for anchor in manifest.get("required_anchors", [])}
    legacy_binding_names = {"valuation_input_adjusted_fcf_ttm"}
    for name in list(wb.defined_names):
        if str(name) in current_binding_ids | current_anchor_ids | legacy_binding_names:
            del wb.defined_names[name]
    for entry in bindings:
        if str(entry.get("planning_state") or "active") != "active":
            continue
        sheet_name = str(entry["sheet"])
        if sheet_name not in wb.sheetnames:
            continue
        target = str(entry.get("planner_target") or entry["target"])
        min_col, min_row, _max_col, _max_row = range_boundaries(target)
        coord = absolute_coordinate(wb[sheet_name].cell(min_row, min_col).coordinate)
        wb.defined_names.add(DefinedName(str(entry["binding_id"]), attr_text=f"{quote_sheetname(sheet_name)}!{coord}"))

    for anchor in manifest.get("required_anchors", []):
        sheet_name = str(anchor["sheet"])
        if sheet_name not in wb.sheetnames:
            continue
        explicit_target = str(anchor.get("target") or "")
        if explicit_target:
            min_col, min_row, _max_col, _max_row = range_boundaries(explicit_target)
            cell = wb[sheet_name].cell(min_row, min_col)
        else:
            sheet_def = next(sheet for sheet in manifest["sheets"] if sheet["sheet"] == sheet_name)
            row_idx = 1 if sheet_name in {"QA_Log", "Needs_Review", "QA_Checks"} else _anchor_row_from_zone(sheet_def, anchor["zone_id"])
            cell = wb[sheet_name].cell(row_idx, 1)
        coord = absolute_coordinate(cell.coordinate)
        wb.defined_names.add(DefinedName(str(anchor["anchor_id"]), attr_text=f"{quote_sheetname(sheet_name)}!{coord}"))


def _neutralize_dynamic_headers(wb: Workbook) -> None:
    if "Valuation" in wb.sheetnames:
        ws = wb["Valuation"]
        if isinstance(ws["O7"].value, str) and ws["O7"].value.startswith("Guidance"):
            ws["O7"] = "Guidance"


def _remove_fixed_dimension_validation(wb: Workbook) -> None:
    """Remove the ANF-authored dimension picker from the generic shell."""

    sheet_name = "{ticker}_Investment_Case"
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    retained = []
    for validation in ws.data_validations.dataValidation:
        formula = str(validation.formula1 or "")
        ranges = {str(cell_range) for cell_range in validation.ranges.ranges}
        if "B38" in ranges and formula == '"None,Brand,Geography"':
            continue
        retained.append(validation)
    ws.data_validations.dataValidation = retained


def _ensure_qa_headers(wb: Workbook) -> None:
    for sheet_name, headers in QA_SHEET_HEADERS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        style_source = ws["B1"] if ws["B1"].has_style else ws["A1"]
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=26):
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.value = None
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_idx, header)
            cell.font = copy(style_source.font)
            cell.fill = copy(style_source.fill)
            cell.border = copy(style_source.border)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.number_format = copy(style_source.number_format)
            ws.column_dimensions[get_column_letter(col_idx)].width = QA_COLUMN_WIDTHS[sheet_name][col_idx - 1]
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"


def _apply_neutral_static_labels(wb: Workbook) -> None:
    for sheet_name, replacements in NEUTRAL_STATIC_LABELS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for coordinate, label in replacements.items():
            cell = ws[coordinate]
            if not isinstance(cell, MergedCell):
                cell.value = label


def _clear_visible_placeholder_labels(wb: Workbook) -> None:
    placeholder = re.compile(
        r"^\s*\[(?:[^\]]*\bslot(?:\s+\d+)?\b|dimension member\s+\d+|guidance metric[^\]]*|operating driver[^\]]*|quality of earnings item\s+\d+)\](?:\s+(?:sales|sales yoy))?\s*$",
        re.I,
    )
    historical_block = re.compile(r"^\s*\[Historical quarter block\s+(\d+)\]\s*$", re.I)
    generic_dynamic_headers = {
        "[current actual period]": "Current actual period",
        "[current guidance period]": "Current guidance period",
    }
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell) or not isinstance(cell.value, str):
                    continue
                text = cell.value.strip()
                if text.casefold() in generic_dynamic_headers:
                    cell.value = generic_dynamic_headers[text.casefold()]
                    continue
                block_match = historical_block.fullmatch(text)
                if block_match:
                    cell.value = f"Historical quarter notes {block_match.group(1)}"
                    continue
                if placeholder.fullmatch(text):
                    cell.value = None


def _neutralize_blank_bs_signal_fills(wb: Workbook) -> None:
    if "BS_Segments" not in wb.sheetnames:
        return
    ws = wb["BS_Segments"]
    white = PatternFill("solid", fgColor="FFFFFF")
    section_rows = {8, 27, 46, 58, 60, 69, 71}
    for row in range(9, 57):
        if row in section_rows:
            continue
        for column in range(2, 10):
            cell = ws.cell(row, column)
            if cell.value in (None, ""):
                cell.fill = copy(white)


def _configure_summary_liquidity_layout(wb: Workbook) -> None:
    if "SUMMARY" not in wb.sheetnames:
        return
    ws = wb["SUMMARY"]
    if "D45:F45" not in {str(item) for item in ws.merged_cells.ranges}:
        ws.merge_cells("D45:F45")
    ws["D45"] = None
    ws["D45"].alignment = copy(ws["B45"].alignment)
    ws["D45"].font = copy(ws["B45"].font)
    ws["D45"].fill = copy(ws["B45"].fill)
    ws["D45"].border = copy(ws["B45"].border)
    ws["D45"].number_format = "General"


def _remove_company_specific_defined_names(wb: Workbook) -> None:
    for name in list(wb.defined_names):
        defined_name = wb.defined_names[name]
        attr_text = str(getattr(defined_name, "attr_text", "") or "")
        if str(name) in COMPANY_SPECIFIC_DEFINED_NAMES | STALE_UNREFERENCED_DEFINED_NAMES or re.search(r"(?i)\b(?:ANF|A&F|Abercrombie|Hollister)\b", f"{name} {attr_text}"):
            del wb.defined_names[name]


def _ranges_overlap(left: str, right: str) -> bool:
    left_min_col, left_min_row, left_max_col, left_max_row = range_boundaries(left)
    right_min_col, right_min_row, right_max_col, right_max_row = range_boundaries(right)
    return not (
        left_max_col < right_min_col
        or right_max_col < left_min_col
        or left_max_row < right_min_row
        or right_max_row < left_min_row
    )


def _clear_profile_owned_range(wb: Workbook, sheet_name: str, target: str) -> None:
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    for merged_range in list(ws.merged_cells.ranges):
        if _ranges_overlap(str(merged_range), target):
            ws.unmerge_cells(str(merged_range))

    retained_validations = []
    for validation in ws.data_validations.dataValidation:
        if any(_ranges_overlap(str(cell_range), target) for cell_range in validation.ranges.ranges):
            continue
        retained_validations.append(validation)
    ws.data_validations.dataValidation = retained_validations

    retained_rules = {}
    for conditional_range, rules in ws.conditional_formatting._cf_rules.items():
        ranges = getattr(conditional_range, "sqref", ())
        if any(_ranges_overlap(str(cell_range), target) for cell_range in ranges):
            continue
        retained_rules[conditional_range] = rules
    ws.conditional_formatting._cf_rules.clear()
    ws.conditional_formatting._cf_rules.update(retained_rules)

    min_col, min_row, max_col, max_row = range_boundaries(target)
    normal_style = copy(wb._cell_styles[0])  # type: ignore[attr-defined]
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.comment = None
            cell.hyperlink = None
            cell._style = copy(normal_style)  # type: ignore[attr-defined]
            cell.protection = Protection(locked=True)


def _apply_module_profile_boundaries(
    wb: Workbook,
    module_payload: dict[str, Any],
    resolved_profile: ResolvedModuleProfile,
) -> None:
    enabled_modules = set(resolved_profile.enabled_modules)
    selected_packs = set(resolved_profile.profile_pack_ids)
    union_profile_active = resolved_profile.profile_id == str(module_payload["union_shell_profile_id"])
    pack_blocks = [block for block in visible_block_contracts(module_payload) if block.kind == "profile_pack_block"]
    active_pack_slots = {
        (block.sheet, block.target)
        for block in pack_blocks
        if block.owner_id in enabled_modules
        and (union_profile_active or block.source_id in selected_packs)
    }
    cleared: set[tuple[str, str]] = set()
    for block in visible_block_contracts(module_payload):
        disabled = block.owner_id not in enabled_modules
        if block.kind == "profile_pack_block":
            disabled = (block.sheet, block.target) not in active_pack_slots
        key = (block.sheet, block.target)
        if disabled and key not in cleared:
            _clear_profile_owned_range(wb, block.sheet, block.target)
            cleared.add(key)


def _prune_defined_names_for_profile(
    wb: Workbook,
    module_payload: dict[str, Any],
    binding_payload: dict[str, Any],
    resolved_profile: ResolvedModuleProfile,
) -> None:
    allowed = enabled_defined_name_ids(module_payload, binding_payload, resolved_profile)
    for name in list(wb.defined_names):
        if str(name) not in allowed:
            del wb.defined_names[name]


def _remove_qa_excel_tables(wb: Workbook) -> None:
    for sheet_name in ("QA_Log", "Needs_Review", "QA_Checks"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for table_name in list(ws.tables.keys()):
            del ws.tables[table_name]


def _clear_workbook_comments(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.comment is not None:
                    cell.comment = None


def _sheet_refs(text: str) -> set[str]:
    refs: set[str] = set()
    for match in SHEET_REF_RE.finditer(text):
        ref = (match.group(1) or match.group(2) or "").strip()
        if ref:
            refs.add(ref)
    return refs


def _remove_defined_names_pointing_to_missing_sheets(wb: Workbook) -> None:
    sheet_names = set(wb.sheetnames)
    for name in list(wb.defined_names):
        defined_name = wb.defined_names[name]
        text = getattr(defined_name, "attr_text", "") or str(defined_name)
        if any(ref not in sheet_names for ref in _sheet_refs(text)):
            del wb.defined_names[name]


def _neutralize_support_sheet(
    ws: Any,
    headers: list[str],
    *,
    state: str = "hidden",
    protect: bool = False,
) -> None:
    for table_name in list(ws.tables.keys()):
        del ws.tables[table_name]
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if getattr(ws, "data_validations", None) is not None:
        ws.data_validations.dataValidation = []
    if getattr(ws, "conditional_formatting", None) is not None:
        ws.conditional_formatting._cf_rules.clear()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.comment = None
            cell.font = Font(name="Aptos", size=10, color="1F2933")
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.alignment = Alignment(vertical="center")
            cell.number_format = "General"
            cell.protection = Protection(locked=True)
    ws.row_dimensions.clear()
    ws.column_dimensions.clear()
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12.0, min(32.0, len(header) + 2.0))
    ws.row_dimensions[1].height = 20.0
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if ws.title == "History_Q":
        for row in range(2, 1001):
            for column in range(1, len(headers) + 1):
                ws.cell(row, column).protection = Protection(locked=False)
    ws.protection.sheet = protect
    ws.sheet_state = state


def _neutralize_hidden_support_sheets(wb: Workbook, manifest: dict[str, Any]) -> None:
    union_order = [str(sheet_name) for sheet_name in manifest["union_sheet_order"]]
    union_sheets = set(union_order)
    contracts = {str(row["sheet"]): row for row in manifest["sheets"]}
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in union_sheets:
            del wb[sheet_name]

    for sheet_name in union_order:
        contract = contracts[sheet_name]
        if str(contract["module_role"]) == "visible_product":
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Rich shell source lacks required visible sheet {sheet_name!r}.")
            wb[sheet_name].sheet_state = str(contract["state"])
            continue
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        module_contract = next(
            row
            for row in manifest.get("sheets") or []
            if str(row.get("sheet") or "") == sheet_name
        )
        headers = [str(value) for value in module_contract.get("formulas_static_labels") or []]
        _neutralize_support_sheet(
            ws,
            headers,
            state=str(contract["state"]),
            protect=bool(module_contract.get("worksheet_protection")),
        )

    wb._sheets = [wb[sheet_name] for sheet_name in union_order]  # type: ignore[attr-defined]

    _remove_defined_names_pointing_to_missing_sheets(wb)


def _ensure_freeze_panes(wb: Workbook, manifest: dict[str, Any]) -> None:
    for sheet_name in manifest["union_sheet_order"]:
        if sheet_name in wb.sheetnames and not wb[sheet_name].freeze_panes:
            wb[sheet_name].freeze_panes = _fallback_freeze(str(sheet_name))


def _configure_calculation(wb: Workbook) -> None:
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def _configure_deterministic_properties(wb: Workbook) -> None:
    fixed = datetime(2000, 1, 1, 0, 0, 0)
    wb.properties.creator = "Standard Stock Model Template"
    wb.properties.lastModifiedBy = "Standard Stock Model Template"
    wb.properties.created = fixed
    wb.properties.modified = fixed


def _materialize_rich_shell(
    *,
    source_path: Path,
    output_path: Path,
    manifest: dict[str, Any],
    binding_payload: dict[str, Any],
    module_payload: dict[str, Any],
    resolved_profile: ResolvedModuleProfile,
) -> Path:
    bindings = list(binding_payload.get("bindings") or [])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source_path, output_path)
    wb = load_workbook(output_path, data_only=False, read_only=False)
    _rename_investment_case_sheet(wb)
    _hide_nonstandard_sheets(wb, manifest)
    _clear_source_specific_visible_text(wb, manifest)
    _genericize_sector_specific_visible_text(wb, manifest)
    _clear_writable_zones(wb, manifest)
    _clear_investment_case_scenario_surfaces(wb)
    _neutralize_quarter_notes_history(wb)
    _clear_visible_source_values_and_notes(wb, manifest)
    _clear_valuation_numeric_constants(wb)
    _clear_valuation_runtime_value_constants(wb)
    _neutralize_writable_data_like_fills(wb, manifest)
    _neutralize_visible_blank_gray_fills(wb)
    _neutralize_valuation_signal_fills(wb)
    _neutralize_red_green_flags(wb)
    _neutralize_blank_valuation_value_fills(wb)
    _ensure_promise_progress_structure(wb)
    _configure_narrative_text_layout(wb)
    _ensure_valuation_guidance_sidecar_headers(wb)
    _ensure_operating_driver_sheet_headers(wb)
    _neutralize_remaining_visible_row_labels(wb)
    _apply_neutral_static_labels(wb)
    _ensure_sheet_titles(wb, manifest)
    _move_operating_drivers_title_to_row1(wb)
    _neutralize_dynamic_headers(wb)
    _remove_fixed_dimension_validation(wb)
    _ensure_binding_anchor_labels(wb, manifest, bindings)
    _clear_visible_placeholder_labels(wb)
    _neutralize_blank_bs_signal_fills(wb)
    _configure_summary_liquidity_layout(wb)
    apply_standard_formula_contracts(
        wb,
        enabled_formula_ids=enabled_formula_ids(module_payload, resolved_profile),
    )
    _apply_module_profile_boundaries(wb, module_payload, resolved_profile)
    _remove_company_specific_defined_names(wb)
    _neutralize_hidden_support_sheets(wb, manifest)
    apply_standard_support_formula_contracts(
        wb,
        enabled_formula_ids=enabled_formula_ids(module_payload, resolved_profile),
    )
    _prune_defined_names_for_profile(wb, module_payload, binding_payload, resolved_profile)
    _ensure_qa_headers(wb)
    _remove_qa_excel_tables(wb)
    _clear_workbook_comments(wb)
    _ensure_freeze_panes(wb, manifest)
    _configure_calculation(wb)
    _configure_deterministic_properties(wb)
    ownership_issues = validate_workbook_execution_ownership(
        wb,
        module_payload,
        binding_payload,
        resolved_profile,
    )
    if ownership_issues:
        raise ValueError("Invalid materialized module ownership: " + "; ".join(ownership_issues[:20]))
    wb.save(output_path)
    wb.close()
    normalize_xlsx_package(output_path)
    return output_path


def materialize_shell(
    *,
    data_root: Path,
    output_path: Path,
    manifest_path: Path,
    binding_map_path: Path,
    module_manifest_path: Path = DEFAULT_MODULE_MANIFEST,
    module_profile_id: str = "full_union",
    contract_manifest_output_path: Path | None = None,
    contract_binding_map_output_path: Path | None = None,
    update_identity: bool = False,
) -> Path:
    base_manifest = _load_json(manifest_path)
    base_binding_payload = _load_json(binding_map_path)
    module_payload = load_workbook_module_manifest(module_manifest_path)
    resolved_profile = resolve_module_profile(module_payload, module_profile_id)
    manifest = build_profile_shell_manifest(base_manifest, module_payload, resolved_profile)
    _configure_investment_case_ownership_zones(manifest)
    binding_payload = build_profile_binding_payload(base_binding_payload, module_payload, resolved_profile)
    _ensure_hidden_support_planner_contracts(manifest, binding_payload)
    manifest["version"] = "0.3.0"
    manifest["semantic_contract_version"] = SHELL_SEMANTIC_CONTRACT_VERSION
    manifest["formula_contract_version"] = FORMULA_CONTRACT_VERSION
    manifest["optional_support_sheets"] = [
        {
            "sheet": str(row["sheet"]),
            "module_id": str(row["module_id"]),
            "module_role": str(row["module_role"]),
            "state": str(row["state"]),
        }
        for row in manifest["sheets"]
        if str(row["module_role"]) != "visible_product"
    ]
    if (contract_manifest_output_path is None) != (contract_binding_map_output_path is None):
        raise ValueError("Profile contract outputs must provide both manifest and binding-map paths.")
    if module_profile_id != str(module_payload["union_shell_profile_id"]) and contract_manifest_output_path is None:
        raise ValueError("Non-union profile variants require isolated manifest and binding-map output paths.")
    manifest_output = contract_manifest_output_path or manifest_path
    binding_output = contract_binding_map_output_path or binding_map_path
    bindings = list(binding_payload.get("bindings") or [])

    rich_source = _rich_template_source(data_root)
    if rich_source is not None:
        result = _materialize_rich_shell(
            source_path=rich_source,
            output_path=output_path,
            manifest=manifest,
            binding_payload=binding_payload,
            module_payload=module_payload,
            resolved_profile=resolved_profile,
        )
        if update_identity:
            _update_manifest_identity(
                result,
                manifest=manifest,
                binding_payload=binding_payload,
                manifest_output_path=manifest_output,
                binding_output_path=binding_output,
            )
        return result

    contracts = _load_source_contracts(data_root, manifest)

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    wb._standard_template_required_anchors = list(manifest.get("required_anchors") or [])  # type: ignore[attr-defined]

    for sheet_def in manifest["sheets"]:
        sheet_name = str(sheet_def["sheet"])
        ws = wb.create_sheet(sheet_name)
        if str(sheet_def.get("module_role") or "") != "visible_product":
            _neutralize_support_sheet(
                ws,
                [str(value) for value in sheet_def.get("formulas_static_labels") or []],
                state=str(sheet_def["state"]),
                protect=bool(sheet_def.get("worksheet_protection")),
            )
            continue
        sheet_bindings = [entry for entry in bindings if entry["sheet"] == sheet_name]
        _write_static_structure(wb, ws, sheet_def, sheet_bindings, contracts.get(sheet_name, SourceSheetContract(None, {}, {})))
        ws.sheet_state = str(sheet_def["state"])

    _configure_summary_liquidity_layout(wb)
    _configure_narrative_text_layout(wb)
    apply_standard_formula_contracts(
        wb,
        enabled_formula_ids=enabled_formula_ids(module_payload, resolved_profile),
    )
    _apply_module_profile_boundaries(wb, module_payload, resolved_profile)
    apply_standard_support_formula_contracts(
        wb,
        enabled_formula_ids=enabled_formula_ids(module_payload, resolved_profile),
    )
    _prune_defined_names_for_profile(wb, module_payload, binding_payload, resolved_profile)
    _configure_calculation(wb)
    _configure_deterministic_properties(wb)
    ownership_issues = validate_workbook_execution_ownership(
        wb,
        module_payload,
        binding_payload,
        resolved_profile,
    )
    if ownership_issues:
        raise ValueError("Invalid materialized module ownership: " + "; ".join(ownership_issues[:20]))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    normalize_xlsx_package(output_path)
    if update_identity:
        _update_manifest_identity(
            output_path,
            manifest=manifest,
            binding_payload=binding_payload,
            manifest_output_path=manifest_output,
            binding_output_path=binding_output,
        )
    return output_path


def _range_contains(container: str, target: str) -> bool:
    left, top, right, bottom = range_boundaries(container)
    target_left, target_top, target_right, target_bottom = range_boundaries(target)
    return left <= target_left and top <= target_top and right >= target_right and bottom >= target_bottom


def _ensure_hidden_support_planner_contracts(
    manifest: dict[str, Any],
    binding_payload: dict[str, Any],
) -> None:
    """Add exact contracts for active table bindings on declared support capacity."""

    support_rows = {
        str(row.get("sheet") or ""): row
        for row in manifest.get("sheets") or []
        if str(row.get("module_role") or "") != "visible_product"
    }
    support_sheets = set(support_rows)
    zone_ids_by_sheet: dict[str, set[str]] = {}
    for binding in binding_payload.get("bindings") or []:
        sheet = str(binding.get("sheet") or "")
        if (
            sheet in support_sheets
            and str(binding.get("planning_state") or "active") == "active"
            and binding.get("writable") is True
        ):
            zone_ids_by_sheet.setdefault(sheet, set()).add(str(binding.get("shell_zone") or ""))
    for sheet, zone_ids in zone_ids_by_sheet.items():
        if len(zone_ids) != 1 or "" in zone_ids:
            raise ValueError(f"{sheet}: active support bindings require one shared non-empty shell zone.")
        writable_zones = support_rows[sheet].get("writable_zones") or []
        if len(writable_zones) != 1:
            raise ValueError(f"{sheet}: support binding requires exactly one writable shell zone.")
        writable_zones[0]["zone_id"] = next(iter(zone_ids))
    contracts = list(manifest.get("planner_cell_contracts") or [])
    for binding in binding_payload.get("bindings") or []:
        if (
            str(binding.get("sheet") or "") not in support_sheets
            or str(binding.get("planning_state") or "active") != "active"
            or binding.get("writable") is not True
            or str(binding.get("planning_mode") or "") not in {"table_rows", "validation_rows"}
        ):
            continue
        binding_id = str(binding.get("binding_id") or "")
        sheet = str(binding.get("sheet") or "")
        planner_target = str(binding.get("planner_target") or binding.get("target") or "")
        left, top, right, bottom = range_boundaries(planner_target)
        for column in binding.get("target_columns") or []:
            target_column = str(column.get("target_column") or "").upper()
            role = str(column.get("target_role") or "")
            target_type = str(column.get("target_type") or "")
            if not role or not target_type:
                raise ValueError(f"{binding_id}: support target {target_column} requires target_role and target_type.")
            column_index = range_boundaries(f"{target_column}1")[0]
            if column_index < left or column_index > right:
                raise ValueError(f"{binding_id}: target column {target_column} is outside {planner_target}.")
            target = f"{target_column}{top}:{target_column}{bottom}"
            same_sheet = [row for row in contracts if str(row.get("sheet") or "") == sheet]
            matches = [row for row in same_sheet if _range_contains(str(row.get("target") or ""), target)]
            if len(matches) == 1 and str(matches[0].get("contract_id") or "").startswith("support_"):
                matches[0]["target_role"] = role
                matches[0]["allowed_binding_ids"] = [binding_id]
                matches[0]["allowed_target_types"] = [target_type]
                continue
            if matches:
                continue
            overlaps = [row for row in same_sheet if _ranges_overlap(str(row.get("target") or ""), target)]
            if overlaps:
                raise ValueError(f"{binding_id}: planner target {sheet}!{target} overlaps an existing cell contract.")
            token = re.sub(r"[^a-z0-9]+", "_", f"{binding_id}_{column.get('column_id') or target_column}".lower()).strip("_")
            contracts.append(
                {
                    "contract_id": f"support_{token}",
                    "sheet": sheet,
                    "target": target,
                    "writable": True,
                    "target_role": role,
                    "allowed_binding_ids": [binding_id],
                    "allowed_target_types": [target_type],
                }
            )
    manifest["planner_cell_contracts"] = contracts


def _update_manifest_identity(
    output_path: Path,
    *,
    manifest: dict[str, Any],
    binding_payload: dict[str, Any],
    manifest_output_path: Path,
    binding_output_path: Path,
) -> None:
    manifest["semantic_contract_version"] = SHELL_SEMANTIC_CONTRACT_VERSION
    manifest["shell_identity"] = compute_shell_identity(
        output_path,
        manifest=manifest,
        binding_payload=binding_payload,
        semantic_contract_version=SHELL_SEMANTIC_CONTRACT_VERSION,
    )
    manifest_output_path.parent.mkdir(parents=True, exist_ok=True)
    binding_output_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_output_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    binding_output_path.write_text(json.dumps(binding_payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    if manifest_output_path.resolve() == (ROOT / "docs" / "standard_template_shell_manifest.json").resolve():
        write_audit_freshness(
            shell_path=output_path,
            manifest=manifest,
            binding_payload=binding_payload,
            root=ROOT,
        )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-root", type=Path, default=DEFAULT_DATA_ROOT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--manifest", type=Path, default=ROOT / "docs" / "standard_template_shell_manifest.json")
    parser.add_argument("--binding-map", type=Path, default=ROOT / "docs" / "workbook_binding_map.json")
    parser.add_argument("--module-manifest", type=Path, default=DEFAULT_MODULE_MANIFEST)
    parser.add_argument("--module-profile", default="full_union")
    parser.add_argument("--profile-manifest-output", type=Path, default=None)
    parser.add_argument("--profile-binding-map-output", type=Path, default=None)
    parser.add_argument("--update-identity", action="store_true", help="Update manifest shell_identity after deterministic materialization.")
    args = parser.parse_args()

    path = materialize_shell(
        data_root=args.data_root.expanduser().resolve(),
        output_path=args.output.expanduser().resolve(),
        manifest_path=args.manifest.expanduser().resolve(),
        binding_map_path=args.binding_map.expanduser().resolve(),
        module_manifest_path=args.module_manifest.expanduser().resolve(),
        module_profile_id=str(args.module_profile),
        contract_manifest_output_path=args.profile_manifest_output.expanduser().resolve() if args.profile_manifest_output else None,
        contract_binding_map_output_path=args.profile_binding_map_output.expanduser().resolve() if args.profile_binding_map_output else None,
        update_identity=args.update_identity,
    )
    print(f"standard template shell: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
