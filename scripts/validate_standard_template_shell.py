"""Validate the frozen standard stock-model workbook shell.

This script validates the template artifact only. It does not build, patch, or
validate ticker workbooks and it does not implement the value-only filler
runtime.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from pbi_xbrl.standard_template_shell_identity import (
    _planned_cell_values_equal,
    verify_post_fill_structural_identity,
    verify_shell_identity,
)
from pbi_xbrl.new_ticker_binding_planner import (
    BindingPlanReproductionError,
)
from pbi_xbrl.new_ticker_style_planner import (
    DEFAULT_MODULE_MANIFEST,
    DEFAULT_STYLE_POLICY,
    StylePlanningError,
    load_style_policy_contract,
    reproduce_style_plan,
)
from pbi_xbrl.json_schema_validation import load_json_strict
from pbi_xbrl.excel_formula_serialization import (
    validate_workbook_formula_compatibility,
    validate_xlsx_formula_compatibility,
)
from pbi_xbrl.standard_template_formula_contract import validate_workbook_protection_contract
from pbi_xbrl.workbook_modules import (
    load_workbook_module_manifest,
    resolve_module_profile,
    validate_workbook_execution_ownership,
)

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_standard_template_hidden_support_audit import (  # noqa: E402
    _source_specific_match,
    scan_hidden_support_package,
)
from build_standard_template_shell_neutrality_audit import scan_neutrality_workbook  # noqa: E402
from build_standard_template_sheet_inventory import build_inventory  # noqa: E402


ROOT = REPO_ROOT
DEFAULT_TEMPLATE = ROOT / "templates" / "standard_stock_model_template.xlsx"
DEFAULT_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
DEFAULT_BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
DEFAULT_LAB_SOURCE = ROOT / "templates" / "lab" / "ANF_template_lab.xlsx"
DEFAULT_SHEET_INVENTORY = ROOT / "docs" / "standard_template_sheet_inventory.json"
DEFAULT_SUPPORT_LIFECYCLE = ROOT / "docs" / "support_sheet_lifecycle_contract.json"
RICH_VISIBLE_SHEETS = (
    "SUMMARY",
    "Valuation",
    "BS_Segments",
    "Operating_Drivers",
    "{ticker}_Investment_Case",
    "Quarter_Notes_UI",
    "Promise_Progress_UI",
)

SOURCE_SPECIFIC_TERMS = (
    "ANF",
    "A&F",
    "Pitney Bowes",
    "Green Plains",
    "Abercrombie",
    "Hollister",
    "Presort",
    "SendTech",
    "45Z",
    "RIN",
    "crush margin",
)
REQUIRED_PROMISE_ANNUAL_HEADERS = [
    "Metric",
    "Initial guide",
    "Q1 update",
    "Q2 update",
    "Q3 update",
    "Q4 update",
    "Actual",
    "Status",
    "Notes/source",
]
REQUIRED_PROMISE_REVISION_HEADERS = [
    "Metric",
    "Previous guide",
    "New/current guide",
    "Change type",
    "Actual",
    "Progress / run-rate",
    "Status",
    "Horizon",
    "Stated in",
    "Source date",
    "Source / note",
]
PROMISE_ANNUAL_HEADER_ROWS = (12, 23, 29, 34)
PROMISE_REVISION_HEADER_ROWS = (60, 70, 77, 85, 91, 98)
VALUATION_GUIDANCE_SIDECAR_HEADERS = {
    "O7": "Current guidance",
    "O8": "Metric",
    "Q8": "Stated in",
    "R8": "Applies to",
    "S8": "Guidance",
    "X8": "Unit",
    "Y8": "Published",
    "Z8": "Evidence",
    "AA8": "Role / source status",
    "O27": "Historical guidance",
    "O28": "Metric",
    "Q28": "Stated in",
    "R28": "Applies to",
    "S28": "Guidance",
    "X28": "Unit",
    "Y28": "Published",
    "Z28": "Evidence",
    "AA28": "Role / source status",
    "O48": "Thesis / debate evidence",
    "O49": "Typed evidence only; unresolved synthesis remains explicit.",
    "O50": "Item",
    "Q50": "Evidence",
    "X50": "Review state",
    "Z50": "Source key",
    "O63": "Output",
    "U63": "Value",
    "X63": "Interpretation",
}
OPERATING_DRIVER_SHEET_HEADERS = {
    "A12": "Topic",
    "B12": "Current read",
    "H12": "Source / use",
    "A19": "Horizon",
    "B19": "Stated in",
    "C19": "Commentary",
}
GTX_RE = re.compile(r"\bGTX\b", re.I)
PACKAGE_SOURCE_PATTERNS = tuple(
    re.compile(pattern, re.I)
    for pattern in (
        rb"\bANF\b",
        rb"A&F",
        rb"\bGPRE\b",
        rb"\bGTX\b",
        rb"\bAbercrombie\b",
        rb"\bHollister\b",
        rb"\bPitney Bowes\b",
        rb"\bPresort\b",
        rb"\bSendTech\b",
        rb"\bGreen Plains\b",
        rb"\b45Z\b",
        rb"\bRINs?\b",
        rb"crush margin",
        rb"tickers[\\/]+ANF\b",
        rb"tickers[\\/]+PBI\b",
        rb"tickers[\\/]+GPRE\b",
        rb"tickers[\\/]+GTX\b",
        rb"\banf[-_]",
        rb"\bpbi[-_]",
        rb"\bgpre[-_]",
        rb"\bgtx[-_]",
    )
)

EXPECTED_STATIC_LABELS_BY_SHEET = {
    "SUMMARY": [
        "SUMMARY",
        "What the company does",
        "Current strategic context",
        "Key competitive advantage",
        "Operating model per segment",
        "Key dependencies",
        "What would make me wrong",
        "Key Financials",
        "Leverage / Liquidity",
    ],
    "Valuation": [
        "Scale",
        "Valuation",
        "Actuals",
        "Quarter",
        "Operating",
        "Revenue",
        "EBITDA",
        "Guidance",
        "Metric",
        "Stated in",
        "Applies to",
        "Role / source status",
        "Historical guidance",
        "Thesis / debate evidence",
    ],
    "BS_Segments": [
        "Balance Sheet & Segments",
        "Quarter",
        "Liquidity / Assets",
        "Cash & cash equivalents",
        "Total assets",
        "Total liabilities",
        "Total equity",
        "Shares diluted",
    ],
    "Operating_Drivers": [
        "Operating Drivers",
        "Current watchlist",
        "Watch item",
        "Current/latest outlook",
        "Topic",
    ],
    "{ticker}_Investment_Case": [
        "Investment Snapshot",
        "Key Debate",
        "Typed Scenario Inputs",
        "Scenario Driver Bridge",
    ],
    "Quarter_Notes_UI": [
        "Quarter Notes",
        "Quarter read",
        "Model read",
        "What changed",
        "Watch next",
        "Key caveat",
        "Key developments",
        "Theme",
    ],
    "Promise_Progress_UI": [
        "Promise Progress",
        "Management Credibility Scorecard",
        "Category",
        "Score",
        "Evidence",
        "Read",
        "guidance progression",
    ],
    "QA_Log": ["issue_id", "severity", "rule_id", "issue_type", "section", "occurrence_count", "detail_ref"],
    "Needs_Review": ["issue_id", "severity", "rule_id", "normalized_path", "occurrence_count", "detail_ref"],
    "QA_Checks": ["rule_id", "status", "unique_issue_count", "occurrence_count", "blocking_count", "detail_ref"],
}

MIN_STATIC_TEXT_COUNTS = {
    "SUMMARY": 24,
    "Valuation": 90,
    "BS_Segments": 55,
    "Operating_Drivers": 45,
    "{ticker}_Investment_Case": 95,
    # Internal slot labels are intentionally blank. The exact reusable history
    # block contract below now carries the stronger structure check.
    "Quarter_Notes_UI": 270,
    "Promise_Progress_UI": 40,
    "QA_Log": 10,
    "Needs_Review": 10,
    "QA_Checks": 9,
}

QUARTER_NOTES_HISTORY_BLOCK_ROWS = (32, 65, 96, 127, 158, 189, 218, 247, 274, 301, 328)
QUARTER_NOTES_HISTORY_LABELS = {
    1: "Quarter read",
    2: "Model read",
    3: "What changed",
    4: "Watch next",
    5: "Key caveat",
    7: "Key developments",
}
QUARTER_NOTES_HISTORY_TABLE_HEADERS = {
    1: "Theme",
    3: "What happened",
    6: "Why it matters",
    8: "Model / valuation implication",
    13: "Source / confidence",
}


@dataclass(frozen=True)
class ShellValidationIssue:
    severity: str
    rule_id: str
    sheet: str
    target: str
    message: str

    def to_dict(self) -> dict[str, str]:
        return {
            "severity": self.severity,
            "rule_id": self.rule_id,
            "sheet": self.sheet,
            "target": self.target,
            "message": self.message,
        }


def _load_json(path: Path) -> dict[str, Any]:
    payload = load_json_strict(path)
    if not isinstance(payload, dict):
        raise ValueError(f"JSON contract must be an object: {path}")
    return payload


def _issue(rule_id: str, message: str, *, sheet: str = "", target: str = "", severity: str = "P1") -> ShellValidationIssue:
    return ShellValidationIssue(severity=severity, rule_id=rule_id, sheet=sheet, target=target, message=message)


def _parse_range(target: str) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(target)
    if min_col > max_col or min_row > max_row:
        raise ValueError(f"Invalid reversed range: {target}")
    return int(min_col), int(min_row), int(max_col), int(max_row)


def _overlaps(first: tuple[int, int, int, int], second: tuple[int, int, int, int]) -> bool:
    f_left, f_top, f_right, f_bottom = first
    s_left, s_top, s_right, s_bottom = second
    return not (f_right < s_left or s_right < f_left or f_bottom < s_top or s_bottom < f_top)


def _contains(outer: tuple[int, int, int, int], inner: tuple[int, int, int, int]) -> bool:
    o_left, o_top, o_right, o_bottom = outer
    i_left, i_top, i_right, i_bottom = inner
    return o_left <= i_left and i_right <= o_right and o_top <= i_top and i_bottom <= o_bottom


def _cells_in_ranges(ws: Any, ranges: Iterable[str]) -> Iterable[Any]:
    for range_ref in ranges:
        for row in ws[range_ref]:
            for cell in row:
                yield cell


def _text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _sheet_texts(ws: Any) -> list[str]:
    values: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, str):
                values.append(value)
    return values


def _sheet_has_label(ws: Any, label: str) -> bool:
    wanted = _text(label).lower()
    if not wanted:
        return False
    for value in _sheet_texts(ws):
        found = _text(value).lower()
        if found == wanted or wanted in found:
            return True
    return False


def _manifest_zone_maps(manifest: dict[str, Any]) -> tuple[dict[tuple[str, str], tuple[int, int, int, int]], dict[str, list[tuple[str, tuple[int, int, int, int]]]]]:
    writable: dict[tuple[str, str], tuple[int, int, int, int]] = {}
    non_writable: dict[str, list[tuple[str, tuple[int, int, int, int]]]] = {}
    for sheet in manifest["sheets"]:
        sheet_name = str(sheet["sheet"])
        for zone in sheet["writable_zones"]:
            writable[(sheet_name, zone["zone_id"])] = _parse_range(str(zone["target"]))
        for zone in sheet["non_writable_zones"]:
            non_writable.setdefault(sheet_name, []).append((zone["zone_id"], _parse_range(str(zone["target"]))))
    return writable, non_writable


def _visible_sheet_names(wb: Any) -> list[str]:
    return [ws.title for ws in wb.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]


def _max_bounds(sheet_def: dict[str, Any]) -> tuple[int, int]:
    max_col = 1
    max_row = 1
    for zone_type in ("writable_zones", "non_writable_zones"):
        for zone in sheet_def[zone_type]:
            min_col, min_row, max_col_in, max_row_in = _parse_range(str(zone["target"]))
            max_col = max(max_col, min_col, max_col_in)
            max_row = max(max_row, min_row, max_row_in)
    return max_col, max_row


def _sheet_name(template: str, ticker: str = "ANF") -> str:
    return template.replace("{ticker}", ticker)


def _filled_ticker_sheet_name(wb: Any) -> str | None:
    return _filled_ticker_sheet_names(wb).get("{ticker}_Investment_Case")


def _filled_ticker_sheet_names(wb: Any) -> dict[str, str]:
    candidates = [
        name
        for name in wb.sheetnames
        if name.endswith("_Investment_Case") and name != "{ticker}_Investment_Case"
    ]
    if len(candidates) != 1:
        return {}
    main = candidates[0]
    ticker = main[: -len("_Investment_Case")]
    resolved = {"{ticker}_Investment_Case": main}
    data_sheet = f"{ticker}_Investment_Case_Data"
    if data_sheet in wb.sheetnames:
        resolved["{ticker}_Investment_Case_Data"] = data_sheet
    return resolved


def _workbook_sheet_name(wb: Any, template_sheet_name: str, *, allow_filled_values: bool) -> str:
    if template_sheet_name in wb.sheetnames:
        return template_sheet_name
    if allow_filled_values:
        resolved = _filled_ticker_sheet_names(wb).get(template_sheet_name)
        if resolved:
            return resolved
    return template_sheet_name


def _expected_visible_sheets(wb: Any, expected_visible: list[str], *, allow_filled_values: bool) -> list[str]:
    if not allow_filled_values:
        return expected_visible
    resolved = _filled_ticker_sheet_name(wb)
    if not resolved:
        return expected_visible
    return [resolved if name == "{ticker}_Investment_Case" else name for name in expected_visible]


def _sheet_counts(wb: Any, sheet_name: str) -> tuple[int, int, int]:
    ws = wb[sheet_name]
    nonempty = 0
    formulas = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            nonempty += 1
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas += 1
    return len(ws.merged_cells.ranges), nonempty, formulas


def _blank_cell_count(ws: Any, target: str) -> int:
    min_col, min_row, max_col, max_row = _parse_range(target)
    count = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value not in (None, ""):
                count += 1
    return count


def _target_row_has_label(ws: Any, target: str) -> bool:
    _min_col, min_row, _max_col, _max_row = _parse_range(target)
    for col_idx in range(1, min(ws.max_column, 3) + 1):
        if _text(ws.cell(min_row, col_idx).value):
            return True
    return False


def _package_source_leakage_parts(path: Path) -> list[dict[str, str]]:
    hits: list[dict[str, str]] = []
    with zipfile.ZipFile(path) as zf:
        for name in zf.namelist():
            data = zf.read(name)
            for pattern in PACKAGE_SOURCE_PATTERNS:
                match = pattern.search(data)
                if not match:
                    continue
                start = max(0, match.start() - 80)
                end = min(len(data), match.end() + 80)
                sample = data[start:end].decode("utf-8", errors="replace")
                hits.append(
                    {
                        "part": name,
                        "pattern": pattern.pattern.decode("ascii", errors="replace"),
                        "sample": re.sub(r"\s+", " ", sample).strip(),
                    }
                )
                break
    return hits


def _unplanned_hidden_company_source_text(
    wb: Any,
    planned_values: dict[tuple[str, str], Any],
) -> dict[str, list[str]]:
    hits: dict[str, list[str]] = {}
    missing = object()
    for ws in wb.worksheets:
        if ws.sheet_state == "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                if not _source_specific_match(cell.value):
                    continue
                expected = planned_values.get((ws.title, cell.coordinate), missing)
                if expected is not missing and _planned_cell_values_equal(cell.value, expected):
                    continue
                hits.setdefault(ws.title, []).append(f"{cell.coordinate}={str(cell.value)[:160]}")
    return hits


def validate_shell(
    *,
    template_path: Path = DEFAULT_TEMPLATE,
    manifest_path: Path = DEFAULT_MANIFEST,
    binding_map_path: Path = DEFAULT_BINDING_MAP,
    module_manifest_path: Path = DEFAULT_MODULE_MANIFEST,
    style_policy_path: Path = DEFAULT_STYLE_POLICY,
    allow_filled_values: bool = False,
    approved_shell_path: Path = DEFAULT_TEMPLATE,
    approved_plan_path: Path | None = None,
    normalized_package_path: Path | None = None,
) -> dict[str, Any]:
    issues: list[ShellValidationIssue] = []
    approved_planned_values: dict[tuple[str, str], Any] = {}
    reproduced_plan = None
    if not template_path.exists():
        return {
            "status": "FAIL",
            "template_path": str(template_path),
            "issues": [_issue("template_missing", "Workbook file does not exist.").to_dict()],
        }
    if template_path.suffix.lower() != ".xlsx":
        issues.append(_issue("template_file_type", "Template must be a macro-free .xlsx file.", target=str(template_path)))
    if template_path.suffix.lower() == ".xlsm":
        issues.append(_issue("template_macro_file_type", "Template must not be .xlsm.", target=str(template_path)))
    try:
        with zipfile.ZipFile(template_path) as zf:
            names = [name.lower() for name in zf.namelist()]
            if any(name.endswith("vbaproject.bin") for name in names):
                issues.append(_issue("template_contains_macros", "Template archive contains vbaProject.bin."))
    except zipfile.BadZipFile:
        issues.append(_issue("template_bad_zip", "Template is not a valid xlsx zip archive."))

    if not any(issue.rule_id == "template_bad_zip" for issue in issues):
        neutrality_package_path = approved_shell_path if allow_filled_values else template_path
        try:
            for hit in _package_source_leakage_parts(neutrality_package_path):
                issues.append(
                    _issue(
                        "package_company_source_leakage",
                        "Workbook package contains company/source-specific text outside the normalized package.",
                        target=f"{hit['part']} | {hit['sample']}",
                    )
                )
        except zipfile.BadZipFile:
            pass

    manifest = _load_json(manifest_path)
    binding_payload = _load_json(binding_map_path)
    bindings = list(binding_payload.get("bindings") or [])
    module_payload = _load_json(module_manifest_path)
    try:
        style_contract = load_style_policy_contract(
            style_policy_path,
            module_payload=module_payload,
            binding_payload=binding_payload,
        )
    except StylePlanningError as exc:
        style_contract = None
        issues.append(_issue("style_policy_contract", str(exc), target=str(style_policy_path)))
    identity_report: dict[str, Any] = {}
    if not allow_filled_values:
        verified_identity = verify_shell_identity(
            template_path,
            manifest=manifest,
            binding_payload=binding_payload,
        )
        identity_report = verified_identity.to_dict()
        for identity_issue in identity_report.get("issues") or []:
            issues.append(
                _issue(
                    str(identity_issue.get("rule_id") or "shell_identity_failure"),
                    str(identity_issue.get("message") or "Shell identity verification failed."),
                    target=str(template_path),
                )
            )
    else:
        approved_plan = _load_json(approved_plan_path) if approved_plan_path is not None else None
        normalized_package = _load_json(normalized_package_path) if normalized_package_path is not None else None
        reproduced_style_plan = None
        if approved_plan is None or normalized_package is None:
            issues.append(
                _issue(
                    "post_fill_reproduction_inputs_missing",
                    "Filled-workbook validation requires both the normalized package and serialized audit plan.",
                )
            )
        else:
            try:
                reproduced_plan, reproduced_style_plan = reproduce_style_plan(
                    normalized_package,
                    binding_payload=binding_payload,
                    manifest=manifest,
                    shell_path=approved_shell_path,
                    module_payload=module_payload,
                    style_contract=style_contract,
                    expected_binding_plan=approved_plan,
                )
                approved_planned_values = {
                    (str(write["target_sheet"]), str(write["target_cell"])): write.get("value")
                    for write in reproduced_plan.to_dict().get("planned_writes") or []
                }
            except (BindingPlanReproductionError, StylePlanningError) as exc:
                issues.append(_issue("post_fill_plan_reproduction_failed", str(exc)))
        post_fill_identity = verify_post_fill_structural_identity(
            template_path,
            approved_shell_path=approved_shell_path,
            manifest=manifest,
            binding_payload=binding_payload,
            approved_plan=reproduced_plan,
            normalized_package=normalized_package,
            module_payload=module_payload,
            style_contract=style_contract,
            approved_style_plan=reproduced_style_plan,
        )
        identity_report = post_fill_identity
        for identity_issue in post_fill_identity.get("issues") or []:
            issues.append(
                _issue(
                    str(identity_issue.get("rule_id") or "post_fill_structural_identity_failure"),
                    str(identity_issue.get("message") or "Filled workbook structural identity verification failed."),
                    target=str(template_path),
                )
            )
    writable_zones, non_writable_zones = _manifest_zone_maps(manifest)

    for sheet in manifest["sheets"]:
        sheet_name = str(sheet["sheet"])
        for zone_type in ("writable_zones", "non_writable_zones"):
            for zone in sheet[zone_type]:
                try:
                    _parse_range(str(zone["target"]))
                except ValueError as exc:
                    issues.append(_issue("manifest_invalid_range", str(exc), sheet=sheet_name, target=str(zone.get("target"))))
        for writable in sheet["writable_zones"]:
            w_range = _parse_range(str(writable["target"]))
            for non_writable in sheet["non_writable_zones"]:
                nw_range = _parse_range(str(non_writable["target"]))
                if _overlaps(w_range, nw_range):
                    issues.append(
                        _issue(
                            "manifest_zone_overlap",
                            f"Writable zone {writable['zone_id']} overlaps non-writable zone {non_writable['zone_id']}.",
                            sheet=sheet_name,
                            target=str(writable["target"]),
                        )
                    )

    try:
        wb = load_workbook(template_path, data_only=False, read_only=False)
    except Exception as exc:
        return {
            "status": "FAIL",
            "template_path": str(template_path),
            "issues": [issue.to_dict() for issue in [*issues, _issue("template_load_failed", str(exc))]],
        }

    enabled_formula_ids = (
        manifest.get("module_profile", {}).get("enabled_formula_ids", ())
        if isinstance(manifest.get("module_profile"), dict)
        else ()
    )
    for compatibility_issue in validate_workbook_formula_compatibility(wb):
        issues.append(
            _issue(
                str(compatibility_issue.get("rule_id") or "formula_compatibility_failure"),
                str(compatibility_issue.get("message") or "Workbook formula compatibility failed."),
            )
        )
    try:
        xml_formula_issues = validate_xlsx_formula_compatibility(template_path)
    except Exception as exc:
        xml_formula_issues = [{"rule_id": "formula_xml_inventory_unavailable", "message": str(exc)}]
    for compatibility_issue in xml_formula_issues:
        issues.append(
            _issue(
                str(compatibility_issue.get("rule_id") or "formula_xml_compatibility_failure"),
                str(compatibility_issue.get("message") or "Workbook XML formula compatibility failed."),
            )
        )
    for protection_issue in validate_workbook_protection_contract(wb, enabled_formula_ids):
        issues.append(
            _issue(
                str(protection_issue.get("rule_id") or "workbook_protection_failure"),
                str(protection_issue.get("message") or "Workbook protection contract failed."),
                sheet=str(protection_issue.get("sheet") or ""),
                target=str(protection_issue.get("target") or ""),
            )
        )

    expected_visible = _expected_visible_sheets(
        wb,
        list(manifest["visible_sheet_order"]),
        allow_filled_values=allow_filled_values,
    )
    if allow_filled_values and reproduced_plan is not None:
        desired_visible = set(expected_visible)
        for sheet_name, state in reproduced_plan.sheet_visibility.items():
            workbook_sheet = _workbook_sheet_name(wb, sheet_name, allow_filled_values=True)
            if state == "visible":
                desired_visible.add(workbook_sheet)
            else:
                desired_visible.discard(workbook_sheet)
        manifest_order = [
            _workbook_sheet_name(wb, str(sheet["sheet"]), allow_filled_values=True)
            for sheet in manifest["sheets"]
        ]
        expected_visible = [sheet_name for sheet_name in manifest_order if sheet_name in desired_visible]
    actual_visible = _visible_sheet_names(wb)
    if actual_visible != expected_visible:
        issues.append(
            _issue(
                "visible_sheet_order",
                f"Visible sheet order mismatch. expected={expected_visible!r} actual={actual_visible!r}",
            )
        )
    if "{ticker}_Investment_Case" not in wb.sheetnames and not (allow_filled_values and _filled_ticker_sheet_name(wb)):
        issues.append(_issue("ticker_token_sheet_missing", "Tokenized investment-case sheet is missing."))
    filled_ticker_sheet = _filled_ticker_sheet_name(wb) if allow_filled_values else None
    for forbidden in ("PBI_Investment_Case", "GPRE_Investment_Case", "ANF_Investment_Case", "GTX_Investment_Case"):
        if allow_filled_values and forbidden == filled_ticker_sheet:
            continue
        if forbidden in wb.sheetnames:
            issues.append(_issue("ticker_specific_sheet_name", f"Template contains ticker-specific sheet {forbidden}.", sheet=forbidden))

    neutrality_template_path = approved_shell_path if allow_filled_values else template_path
    hidden_audit = scan_hidden_support_package(
        template_path=neutrality_template_path,
        lab_path=DEFAULT_LAB_SOURCE,
        manifest_path=manifest_path,
    )
    package_check = hidden_audit["package_dependency_check"]
    for item in package_check["missing_visible_formula_sheets"]:
        issues.append(_issue("missing_visible_formula_sheet_ref", "Visible formula references a missing sheet.", target=item))
    for item in package_check["missing_defined_name_sheets"]:
        issues.append(_issue("missing_defined_name_sheet_ref", "Defined name references a missing sheet.", target=item))
    for item in package_check["missing_data_validation_sheets"]:
        issues.append(_issue("missing_data_validation_sheet_ref", "Data validation references a missing sheet.", target=item))
    retained_unclassified = hidden_audit["post_neutralization_summary"]["retained_unclassified_hidden_sheets"]
    for sheet_name in retained_unclassified:
        issues.append(_issue("unclassified_hidden_sheet_retained", "Hidden sheet remains without an allowed shell classification.", sheet=sheet_name))
    unplanned_hidden_leakage = (
        _unplanned_hidden_company_source_text(wb, approved_planned_values)
        if allow_filled_values
        else {}
    )
    hidden_leakage = (
        sum(len(samples) for samples in unplanned_hidden_leakage.values())
        if allow_filled_values
        else int(hidden_audit["post_neutralization_summary"]["company_source_leakage_cells"])
    )
    if hidden_leakage:
        issues.append(
            _issue(
                "hidden_company_source_leakage",
                f"Hidden workbook package contains company/source-specific leakage cells. count={hidden_leakage}",
            )
        )
    for sheet_name, samples in unplanned_hidden_leakage.items():
        issues.append(
            _issue(
                "hidden_sheet_company_source_text",
                "Hidden sheet contains unplanned ANF/PBI/GPRE/GTX company/source-specific text.",
                sheet=sheet_name,
                target="; ".join(samples[:3]),
            )
        )
    for row in hidden_audit["hidden_support_sheets"]:
        if not row["present_in_shell"]:
            continue
        if not allow_filled_values and row["contains_company_source_text"]:
            issues.append(
                _issue(
                    "hidden_sheet_company_source_text",
                    "Hidden sheet contains ANF/PBI/GPRE/GTX company/source-specific text.",
                    sheet=row["sheet_name"],
                    target="; ".join(row["company_source_leakage_samples"][:3]),
                )
            )
        if row["contains_source_raw_audit_data"] and row["classification"] not in {
            "keep_formula_dependency",
            "keep_neutral_helper_shell",
            "keep_optional_runtime_output_shell",
        }:
            issues.append(
                _issue(
                    "hidden_raw_source_data_retained",
                    "Hidden raw/source/audit sheet data remains in the frozen shell package.",
                    sheet=row["sheet_name"],
                )
            )

    neutrality_audit = scan_neutrality_workbook(
        template_path=neutrality_template_path,
        manifest_path=manifest_path,
    )
    neutrality_summary = neutrality_audit["post_neutrality_summary"]
    neutrality_rule_map = {
        "company_specific_value_count": "neutrality_company_specific_values",
        "company_specific_text_count": "neutrality_company_specific_text",
        "sector_specific_label_count": "neutrality_sector_specific_labels",
        "fixed_dimension_member_count": "neutrality_fixed_dimension_members",
        "source_specific_text_count": "neutrality_source_specific_text",
        "valuation_numeric_constant_count": "neutrality_valuation_numeric_constants",
        "signal_fill_without_value_count": "neutrality_signal_fills_without_data",
        "blank_writable_non_neutral_fill_count": "neutrality_blank_writable_non_neutral_fills",
        "visible_blank_gray_fill_count": "neutrality_visible_blank_gray_fills",
        "valuation_signal_fill_count": "valuation_signal_fills_without_data",
        "blank_status_or_value_fill_count": "blank_status_or_value_fills",
        "red_green_status_output_count": "red_green_status_outputs",
        "visible_value_date_status_constant_count": "visible_value_date_status_constants",
        "visible_company_source_text_count": "visible_company_source_text",
        "missing_required_support_shell_sheet_count": "missing_required_support_shell_sheets",
    }
    neutrality_samples = [
        *neutrality_audit["non_neutral_items"],
        *neutrality_audit["style_signal_items"],
        *neutrality_audit.get("gray_fill_items", []),
        *neutrality_audit.get("visible_gray_fill_items", []),
        *neutrality_audit.get("valuation_signal_items", []),
        *neutrality_audit.get("blank_status_fill_items", []),
        *neutrality_audit.get("red_green_status_items", []),
        *neutrality_audit.get("visible_value_date_status_constant_items", []),
    ]
    sample_target = "; ".join(
        f"{item['sheet']}!{item['cell']}" for item in neutrality_samples[:5]
    )
    for key, rule_id in neutrality_rule_map.items():
        count = int(neutrality_summary.get(key, 0))
        if count:
            issues.append(
                _issue(
                    rule_id,
                    f"Standard shell contains non-neutral template content. {key}={count}",
                    target=sample_target,
                )
            )

    if not allow_filled_values:
        try:
            computed_inventory, computed_lifecycle = build_inventory(
                template_path=template_path,
                manifest_path=manifest_path,
                binding_map_path=binding_map_path,
            )
            if not DEFAULT_SHEET_INVENTORY.exists():
                issues.append(_issue("sheet_inventory_missing", "Sheet inventory report is missing.", target=str(DEFAULT_SHEET_INVENTORY)))
            else:
                recorded_inventory = _load_json(DEFAULT_SHEET_INVENTORY)
                if len(recorded_inventory.get("sheets", [])) != len(computed_inventory["sheets"]):
                    issues.append(
                        _issue(
                            "sheet_inventory_stale",
                            "Sheet inventory report row count does not match current shell/source inventory.",
                            target=str(DEFAULT_SHEET_INVENTORY),
                        )
                    )
            if not DEFAULT_SUPPORT_LIFECYCLE.exists():
                issues.append(_issue("support_lifecycle_missing", "Support-sheet lifecycle contract is missing.", target=str(DEFAULT_SUPPORT_LIFECYCLE)))
            else:
                recorded_lifecycle = _load_json(DEFAULT_SUPPORT_LIFECYCLE)
                lifecycle_names = {str(row["sheet_name"]) for row in recorded_lifecycle.get("support_sheets", [])}
                computed_names = {str(row["sheet_name"]) for row in computed_lifecycle["support_sheets"]}
                missing_lifecycle = sorted(computed_names - lifecycle_names)
                for sheet_name in missing_lifecycle:
                    issues.append(
                        _issue(
                            "support_lifecycle_sheet_missing",
                            "Support/audit sheet is missing from lifecycle contract.",
                            sheet=sheet_name,
                            target=str(DEFAULT_SUPPORT_LIFECYCLE),
                        )
                    )
        except Exception as exc:
            issues.append(_issue("sheet_inventory_lifecycle_check_failed", str(exc)))

    if DEFAULT_LAB_SOURCE.exists():
        lab_wb = load_workbook(DEFAULT_LAB_SOURCE, data_only=False, read_only=False)
        try:
            for sheet_name in RICH_VISIBLE_SHEETS:
                workbook_sheet = _workbook_sheet_name(wb, sheet_name, allow_filled_values=allow_filled_values)
                if workbook_sheet not in wb.sheetnames or _sheet_name(sheet_name) not in lab_wb.sheetnames:
                    continue
                template_merges, template_nonempty, template_formulas = _sheet_counts(wb, workbook_sheet)
                lab_merges, lab_nonempty, lab_formulas = _sheet_counts(lab_wb, _sheet_name(sheet_name))
                if template_merges < max(1, int(lab_merges * 0.55)):
                    issues.append(
                        _issue(
                            "rich_shell_merge_family_sparse",
                            f"Template merge family is too sparse versus ANF lab. template={template_merges} lab={lab_merges}",
                            sheet=sheet_name,
                        )
                    )
                if template_nonempty < max(5, int(lab_nonempty * 0.03)):
                    issues.append(
                        _issue(
                            "rich_shell_static_content_sparse",
                            f"Template static content is too sparse versus ANF lab. template={template_nonempty} lab={lab_nonempty}",
                            sheet=sheet_name,
                        )
                    )
                if lab_formulas and template_formulas < max(1, int(lab_formulas * 0.04)):
                    issues.append(
                        _issue(
                            "rich_shell_formula_family_sparse",
                            f"Template formula family is too sparse versus ANF lab. template={template_formulas} lab={lab_formulas}",
                            sheet=sheet_name,
                        )
                    )
        finally:
            lab_wb.close()

    for sheet in manifest["sheets"]:
        sheet_name = str(sheet["sheet"])
        workbook_sheet_name = _workbook_sheet_name(wb, sheet_name, allow_filled_values=allow_filled_values)
        if workbook_sheet_name not in wb.sheetnames:
            issues.append(_issue("manifest_sheet_missing", "Manifest sheet does not exist in workbook.", sheet=sheet_name))
            continue
        ws = wb[workbook_sheet_name]
        if (
            ws.sheet_state == "visible"
            and sheet_name in set(manifest["visible_sheet_order"])
            and sheet_name not in {"QA_Log", "Needs_Review", "QA_Checks"}
            and not ws.merged_cells.ranges
        ):
            issues.append(_issue("sheet_merge_contract", "Non-QA visible sheet should retain title/section merges.", sheet=sheet_name))
        if sheet_name in {"QA_Log", "Needs_Review", "QA_Checks"} and ws.tables:
            issues.append(
                _issue(
                    "stale_qa_table_definition",
                    "QA output sheets must not retain source workbook Excel Table definitions.",
                    sheet=sheet_name,
                    target=",".join(ws.tables.keys()),
                )
            )
        if not ws.freeze_panes:
            issues.append(_issue("sheet_freeze_panes_missing", "Sheet freeze panes are missing.", sheet=sheet_name))
        if not any(cell.font and bool(cell.font.bold) for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)) for cell in row):
            issues.append(_issue("sheet_style_contract", "Sheet does not contain a styled/bold header area.", sheet=sheet_name))
        if sheet_name == "Valuation":
            for coord, expected in VALUATION_GUIDANCE_SIDECAR_HEADERS.items():
                if ws[coord].value != expected:
                    issues.append(
                        _issue(
                            "valuation_guidance_sidecar_header_missing",
                            f"Valuation guidance sidecar header is missing or changed. expected={expected!r}",
                            sheet=sheet_name,
                            target=coord,
                        )
                    )
        if sheet_name == "Operating_Drivers":
            for coord, expected in OPERATING_DRIVER_SHEET_HEADERS.items():
                if ws[coord].value != expected:
                    issues.append(
                        _issue(
                            "operating_drivers_standard_subheader_missing",
                            f"Operating_Drivers standard subheader is missing or changed. expected={expected!r}",
                            sheet=sheet_name,
                            target=coord,
                        )
                    )
        if sheet_name == "Quarter_Notes_UI":
            merge_contract = {str(item) for item in ws.merged_cells.ranges}
            for block_index, row_idx in enumerate(QUARTER_NOTES_HISTORY_BLOCK_ROWS, start=1):
                expected_title = f"Historical quarter notes {block_index}"
                if ws.cell(row_idx, 1).value != expected_title:
                    issues.append(
                        _issue(
                            "quarter_notes_history_block_title_missing",
                            f"Historical Quarter Notes block title is missing or changed. expected={expected_title!r}",
                            sheet=sheet_name,
                            target=f"A{row_idx}",
                        )
                    )
                expected_merge = f"A{row_idx}:O{row_idx}"
                if expected_merge not in merge_contract:
                    issues.append(
                        _issue(
                            "quarter_notes_history_block_merge_missing",
                            "Historical Quarter Notes block title merge is missing or changed.",
                            sheet=sheet_name,
                            target=expected_merge,
                        )
                    )
                for offset, expected_label in QUARTER_NOTES_HISTORY_LABELS.items():
                    label_row = row_idx + offset
                    if ws.cell(label_row, 1).value != expected_label:
                        issues.append(
                            _issue(
                                "quarter_notes_history_block_structure_missing",
                                f"Historical Quarter Notes reusable block label is missing or changed. expected={expected_label!r}",
                                sheet=sheet_name,
                                target=f"A{label_row}",
                            )
                        )
                table_header_row = row_idx + 8
                for column, expected_label in QUARTER_NOTES_HISTORY_TABLE_HEADERS.items():
                    if ws.cell(table_header_row, column).value != expected_label:
                        issues.append(
                            _issue(
                                "quarter_notes_history_table_header_missing",
                                f"Historical Quarter Notes table header is missing or changed. expected={expected_label!r}",
                                sheet=sheet_name,
                                target=ws.cell(table_header_row, column).coordinate,
                            )
                        )
        if sheet_name == "Promise_Progress_UI":
            for row_idx in PROMISE_ANNUAL_HEADER_ROWS:
                annual_headers = [ws.cell(row_idx, col).value for col in range(1, 10)]
                if annual_headers != REQUIRED_PROMISE_ANNUAL_HEADERS:
                    issues.append(
                        _issue(
                            "promise_progress_annual_headers_missing",
                            "Promise Progress annual guidance table is missing required standard columns.",
                            sheet=sheet_name,
                            target=f"A{row_idx}:I{row_idx}",
                        )
                    )
            for row_idx in PROMISE_REVISION_HEADER_ROWS:
                revision_headers = [ws.cell(row_idx, col).value for col in range(1, 12)]
                if revision_headers != REQUIRED_PROMISE_REVISION_HEADERS:
                    issues.append(
                        _issue(
                            "promise_progress_revision_headers_missing",
                            "Promise Progress guidance revision table is missing required standard columns.",
                            sheet=sheet_name,
                            target=f"A{row_idx}:K{row_idx}",
                        )
                    )

        writable_ranges = [str(zone["target"]) for zone in sheet["writable_zones"]]
        static_text_count = 0
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.strip() and not value.startswith("="):
                    static_text_count += 1
                    if not allow_filled_values:
                        for term in SOURCE_SPECIFIC_TERMS:
                            if re.search(r"\b" + re.escape(term) + r"\b", value, re.I):
                                issues.append(
                                    _issue(
                                        "source_specific_text_in_visible_template",
                                        f"Company/sector-specific text remains in visible shell: {term}.",
                                        sheet=sheet_name,
                                        target=cell.coordinate,
                                    )
                                )
                                break
        min_static_text = MIN_STATIC_TEXT_COUNTS.get(sheet_name)
        if min_static_text is not None and static_text_count < min_static_text:
            issues.append(
                _issue(
                    "rich_shell_static_label_count_sparse",
                    f"Static/template label count is too low. actual={static_text_count} minimum={min_static_text}",
                    sheet=sheet_name,
                )
            )
        for expected_label in EXPECTED_STATIC_LABELS_BY_SHEET.get(sheet_name, []):
            if not _sheet_has_label(ws, expected_label):
                issues.append(
                    _issue(
                        "rich_shell_expected_static_label_missing",
                        f"Expected reusable static/template label is missing: {expected_label!r}",
                        sheet=sheet_name,
                    )
                )

        for writable_range in writable_ranges:
            nonblank = _blank_cell_count(ws, writable_range)
            if nonblank and not allow_filled_values:
                issues.append(
                    _issue(
                        "writable_zone_not_blank",
                        f"Writable shell zone must be blank before value-only filler runs. nonblank_cells={nonblank}",
                        sheet=sheet_name,
                        target=writable_range,
                    )
                )

        for cell in _cells_in_ranges(ws, writable_ranges):
            value = cell.value
            if value is None or value == "":
                continue
            if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                issues.append(_issue("formula_inside_writable_zone", "Formula found inside writable value zone.", sheet=sheet_name, target=cell.coordinate))
                continue
            value_text = str(value)
            if not allow_filled_values:
                for term in SOURCE_SPECIFIC_TERMS:
                    if re.search(r"\b" + re.escape(term) + r"\b", value_text, re.I):
                        issues.append(
                            _issue(
                                "source_specific_text_in_writable_zone",
                                f"Company/sector-specific text remains in writable zone: {term}.",
                                sheet=sheet_name,
                                target=cell.coordinate,
                            )
                        )
                        break

        for value in _sheet_texts(ws):
            if GTX_RE.search(value):
                issues.append(_issue("gtx_text_present", "GTX-specific text exists in template.", sheet=sheet_name))
                break

    for entry in bindings:
        if not bool(entry.get("writable")):
            continue
        sheet_name = str(entry["sheet"])
        target = str(entry["target"])
        shell_zone = str(entry["shell_zone"])
        target_range = _parse_range(target)
        shell_range = writable_zones.get((sheet_name, shell_zone))
        if shell_range is None:
            issues.append(_issue("binding_shell_zone_missing", f"Binding shell zone {shell_zone} is not declared.", sheet=sheet_name, target=target))
            continue
        if not _contains(shell_range, target_range):
            issues.append(_issue("binding_outside_writable_zone", f"Binding target is outside shell zone {shell_zone}.", sheet=sheet_name, target=target))
        for zone_id, non_writable_range in non_writable_zones.get(sheet_name, []):
            if _overlaps(target_range, non_writable_range):
                issues.append(_issue("binding_overlaps_non_writable_zone", f"Binding target overlaps non-writable zone {zone_id}.", sheet=sheet_name, target=target))
        workbook_sheet_name = _workbook_sheet_name(wb, sheet_name, allow_filled_values=allow_filled_values)
        if workbook_sheet_name in wb.sheetnames:
            label = str(entry.get("anchor_label") or "").strip()
            if label and str(entry.get("binding_id") or "") not in wb.defined_names and not _sheet_has_label(wb[workbook_sheet_name], label) and not _target_row_has_label(wb[workbook_sheet_name], target):
                issues.append(_issue("binding_anchor_label_missing", f"Binding anchor label {label!r} is missing.", sheet=sheet_name, target=target))
            nonblank = _blank_cell_count(wb[workbook_sheet_name], target)
            if nonblank and not allow_filled_values:
                issues.append(
                    _issue(
                        "binding_target_not_blank",
                        f"Writable binding target must be blank in the template. nonblank_cells={nonblank}",
                        sheet=sheet_name,
                        target=target,
                    )
                )

    defined_names = {str(name) for name in wb.defined_names}
    for anchor in manifest.get("required_anchors", []):
        sheet_name = str(anchor["sheet"])
        anchor_id = str(anchor["anchor_id"])
        label = str(anchor.get("anchor_label") or "").strip()
        workbook_sheet_name = _workbook_sheet_name(wb, sheet_name, allow_filled_values=allow_filled_values)
        if anchor_id not in defined_names and (workbook_sheet_name not in wb.sheetnames or not _sheet_has_label(wb[workbook_sheet_name], label)):
            issues.append(_issue("required_anchor_missing", f"Required anchor {anchor_id} / {label!r} is missing.", sheet=sheet_name))

    if not allow_filled_values:
        try:
            module_path = ROOT / str(manifest["module_manifest"]["path"])
            module_payload = load_workbook_module_manifest(module_path)
            resolved_profile = resolve_module_profile(
                module_payload,
                str(manifest["module_profile"]["profile_id"]),
            )
            for ownership_issue in validate_workbook_execution_ownership(
                wb,
                module_payload,
                binding_payload,
                resolved_profile,
            ):
                issues.append(_issue("module_execution_ownership", ownership_issue))
        except (KeyError, TypeError, ValueError) as exc:
            issues.append(_issue("module_execution_contract", str(exc)))

    wb.close()
    status = "PASS" if not issues else "FAIL"
    return {
        "status": status,
        "template_path": str(template_path),
        "issue_count": len(issues),
        "issues": [issue.to_dict() for issue in issues],
        "shell_identity": identity_report,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--binding-map", type=Path, default=DEFAULT_BINDING_MAP)
    parser.add_argument("--module-manifest", type=Path, default=DEFAULT_MODULE_MANIFEST)
    parser.add_argument("--style-policy", type=Path, default=DEFAULT_STYLE_POLICY)
    parser.add_argument("--approved-shell", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--binding-plan", type=Path, help="Exact PASS binding plan required when filled writable values differ from the shell.")
    parser.add_argument("--normalized-package", type=Path, help="Normalized package used to reproduce and authenticate the binding plan.")
    parser.add_argument("--allow-filled-values", action="store_true", help="Validate a filled output workbook layout while allowing mapped values in writable zones.")
    parser.add_argument("--json", action="store_true", help="Print full JSON report.")
    args = parser.parse_args(argv)

    report = validate_shell(
        template_path=args.template.expanduser().resolve(),
        manifest_path=args.manifest.expanduser().resolve(),
        binding_map_path=args.binding_map.expanduser().resolve(),
        module_manifest_path=args.module_manifest.expanduser().resolve(),
        style_policy_path=args.style_policy.expanduser().resolve(),
        allow_filled_values=args.allow_filled_values,
        approved_shell_path=args.approved_shell.expanduser().resolve(),
        approved_plan_path=args.binding_plan.expanduser().resolve() if args.binding_plan else None,
        normalized_package_path=args.normalized_package.expanduser().resolve() if args.normalized_package else None,
    )
    if args.json:
        print(json.dumps(report, indent=2, ensure_ascii=False))
    else:
        print(f"{report['status']}: standard template shell validation")
        print(f"template: {report['template_path']}")
        print(f"issues: {report.get('issue_count', len(report.get('issues', [])))}")
        for issue in report.get("issues", [])[:20]:
            print(f"- {issue['severity']} {issue['rule_id']} {issue['sheet']} {issue['target']}: {issue['message']}")
    return 0 if report["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
