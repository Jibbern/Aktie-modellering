#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import time
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from pbi_xbrl import __version__
from pbi_xbrl.company_profiles import get_company_profile
from pbi_xbrl.cache_layout import bootstrap_canonical_ticker_cache, canonical_ticker_cache_root
from pbi_xbrl.excel_writer import (
    enrich_quarter_notes_audit_rows_with_readback,
    validate_saved_workbook_after_audit_write,
    validate_saved_workbook_export,
    write_quarter_notes_audit_sheet,
)
from pbi_xbrl.market_data import sync_market_cache
from pbi_xbrl.metrics import get_income_statement_rules
from pbi_xbrl.excel_vba import MacroInjectionError, inject_valuation_macros
from pbi_xbrl.pipeline import PipelineConfig, run_pipeline, write_excel
from pbi_xbrl.sec_ingest import IngestConfig, download_all
from pbi_xbrl.sec_xbrl import SecConfig
from pbi_xbrl.xbrl_instance import InstanceMetadata, parse_instance

SEC_USER_AGENT = "Equity model pipeline (contact: Jibbern@outlook.com)"
PIPELINE_BUNDLE_CACHE_VERSION = 2


def _project_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _ticker_root(repo_root: Path, ticker: str | None) -> Path:
    t = str(ticker or "").strip().upper()
    return repo_root / t if t else repo_root


def _excel_output_root(repo_root: Path) -> Path:
    return repo_root / "Excel stock models"


def _require_user_agent(user_agent: str) -> str:
    ua = (user_agent or "").strip()
    if not ua:
        raise SystemExit("ERROR: --user-agent is required (name/org + contact).")
    if "email@example.com" in ua or "YourNameOrOrg" in ua:
        raise SystemExit("ERROR: --user-agent must be your real org/name + contact.")
    return ua


def _default_out_path(ticker: str | None) -> Path:
    base_name = f"{ticker.upper()}_model.xlsm" if ticker else "SEC_data_model.xlsm"
    repo_root = _project_root()
    t = (ticker or "").upper()

    if t:
        return (_excel_output_root(repo_root) / base_name).resolve()

    return (_excel_output_root(repo_root) / base_name).resolve()


def _normalize_out_path_xlsm(path_like: str | Path) -> Path:
    p = Path(path_like).expanduser().resolve()
    if p.suffix.lower() != ".xlsm":
        p = p.with_suffix(".xlsm")
    return p


def _default_step_a_out_path(ticker: str | None) -> Path:
    stem = f"{(ticker or 'SEC_data').upper()}_step_a"
    repo_root = _project_root()
    if ticker:
        return (_ticker_root(repo_root, ticker) / f"{stem}.xlsx").resolve()
    return (_excel_output_root(repo_root) / f"{stem}.xlsx").resolve()


def _normalize_out_path_xlsx(path_like: str | Path) -> Path:
    p = Path(path_like).expanduser().resolve()
    if p.suffix.lower() != ".xlsx":
        p = p.with_suffix(".xlsx")
    return p


def _verify_saved_workbook_export(final_out_path: Path, writer_result: Any) -> Dict[str, Any]:
    if writer_result is None:
        raise SystemExit("ERROR: write_excel did not return a workbook snapshot for Quarter_Notes_UI verification.")
    expected_snapshot = getattr(writer_result, "quarter_notes_ui_snapshot", None)
    if expected_snapshot is None:
        raise SystemExit("ERROR: write_excel result missing Quarter_Notes_UI snapshot for export verification.")
    try:
        return validate_saved_workbook_export(
            final_out_path,
            quarter_notes_ui_snapshot=expected_snapshot,
            summary_export_expectation=getattr(writer_result, "summary_export_expectation", {}) or {},
            valuation_export_expectation=getattr(writer_result, "valuation_export_expectation", {}) or {},
            qa_export_expectation=getattr(writer_result, "qa_export_expectation", None),
            needs_review_export_expectation=getattr(writer_result, "needs_review_export_expectation", None),
        )
    except Exception as exc:
        raise SystemExit(
            f"ERROR: saved workbook verification failed for {final_out_path}: {type(exc).__name__}: {exc}"
        ) from exc


def _hash_file_stats(paths: Iterable[Path], max_files: int = 1200) -> str:
    rows = []
    for p in sorted({Path(x) for x in paths if x is not None}):
        try:
            st = p.stat()
        except Exception:
            continue
        rows.append(f"{p.name}|{int(st.st_size)}|{int(st.st_mtime)}")
        if len(rows) >= max_files:
            break
    if not rows:
        return "none"
    return hashlib.sha1("||".join(rows).encode("utf-8", errors="ignore")).hexdigest()


def _material_signature(repo_root: Path, ticker: Optional[str]) -> str:
    t = str(ticker or "").strip().upper()
    if not t:
        return "none"
    ticker_root = repo_root / t
    if not ticker_root.exists():
        return "missing"
    dirs = [
        ticker_root / "annual_reports",
        ticker_root / "earnings_presentation",
        ticker_root / "earnings_transcripts",
        ticker_root / "financial_statement",
        ticker_root / "press_release",
        ticker_root / f"{t}-10K",
        ticker_root / f"{t}_10K",
        ticker_root / f"{t} 10K",
    ]
    files = []
    for d in dirs:
        if not d.exists() or not d.is_dir():
            continue
        try:
            files.extend([p for p in d.rglob("*") if p.is_file()])
        except Exception:
            continue
    return _hash_file_stats(files, max_files=1500)


def _code_signature(repo_root: Path) -> str:
    files = [
        repo_root / "Code" / "stock_models.py",
        repo_root / "Code" / "pbi_xbrl" / "pipeline.py",
        repo_root / "Code" / "pbi_xbrl" / "metrics.py",
        repo_root / "Code" / "pbi_xbrl" / "non_gaap.py",
        repo_root / "Code" / "pbi_xbrl" / "debt_parser.py",
        repo_root / "Code" / "pbi_xbrl" / "doc_intel.py",
        repo_root / "Code" / "pbi_xbrl" / "guidance_lexicon.py",
        repo_root / "Code" / "pbi_xbrl" / "quarter_notes_lexicon.py",
        repo_root / "Code" / "pbi_xbrl" / "period_resolver.py",
    ]
    return _hash_file_stats(files, max_files=100)


def _default_history_export_path(ticker: str | None, suffix: str) -> Path:
    repo_root = _project_root()
    t = str(ticker or "").strip().upper()
    stem = f"{t}_model_History_Q" if t else f"SEC_data_model_History_Q"
    if t:
        return (_ticker_root(repo_root, t) / f"{stem}{suffix}").resolve()
    return (_excel_output_root(repo_root) / f"{stem}{suffix}").resolve()


def _sec_cache_signature(cache_dir: Path) -> str:
    pats = ["submissions_*.json", "companyfacts_*.json"]
    blobs = []
    for pat in pats:
        try:
            for fp in sorted(cache_dir.glob(pat)):
                try:
                    blobs.append(fp.read_bytes())
                except Exception:
                    st = fp.stat()
                    blobs.append(f"{fp.name}|{int(st.st_size)}".encode("utf-8", errors="ignore"))
        except Exception:
            continue
    if not blobs:
        return "none"
    return hashlib.sha1(b"||".join(blobs)).hexdigest()


def _pipeline_bundle_cache_key(args: argparse.Namespace, cfg: PipelineConfig, repo_root: Path) -> str:
    return "|".join(
        [
            f"v{PIPELINE_BUNDLE_CACHE_VERSION}",
            f"ticker={str(args.ticker or '').upper()}",
            f"cik={str(args.cik or '')}",
            f"max_q={cfg.max_quarters}",
            f"min_year={cfg.min_year}",
            f"tier2={int(cfg.enable_tier2_debt)}",
            f"tier3={int(cfg.enable_tier3_non_gaap)}",
            f"ngaap={cfg.non_gaap_mode}",
            f"strict={cfg.strictness}",
            f"preview={int(cfg.non_gaap_preview)}",
            f"quiet_pdf={int(cfg.quiet_pdf_warnings)}",
            f"skip_doc_intel={int(cfg.use_cached_doc_intel_only)}",
            f"sec={_sec_cache_signature(cfg.cache_dir)}",
            f"materials={_material_signature(repo_root, args.ticker)}",
            f"code={_code_signature(repo_root)}",
        ]
    )


def _pipeline_bundle_cache_paths(cache_dir: Path, ticker: Optional[str]) -> tuple[Path, Path]:
    t = str(ticker or "SEC").upper()
    root = cache_dir / "pipeline_bundle_cache"
    root.mkdir(parents=True, exist_ok=True)
    return root / f"{t}.meta.json", root / f"{t}.pkl"


def _load_pipeline_bundle_cache(cache_dir: Path, ticker: Optional[str], cache_key: str, *, ignore_key: bool = False) -> Optional[Any]:
    meta_path, data_path = _pipeline_bundle_cache_paths(cache_dir, ticker)
    if not (meta_path.exists() and data_path.exists()):
        return None
    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        if not ignore_key and str(meta.get("key")) != str(cache_key):
            return None
        obj = pd.read_pickle(data_path)
        print("[pipeline_bundle_cache] hit", flush=True)
        return obj
    except Exception:
        return None


def _save_pipeline_bundle_cache(cache_dir: Path, ticker: Optional[str], cache_key: str, obj: Any) -> None:
    meta_path, data_path = _pipeline_bundle_cache_paths(cache_dir, ticker)
    try:
        pd.to_pickle(obj, data_path)
        meta_path.write_text(
            json.dumps(
                {
                    "version": PIPELINE_BUNDLE_CACHE_VERSION,
                    "key": cache_key,
                    "saved_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                },
                ensure_ascii=True,
            ),
            encoding="utf-8",
        )
        print("[pipeline_bundle_cache] saved", flush=True)
    except Exception:
        pass


@contextmanager
def _timed(label: str, enabled: bool = True, store: Optional[Dict[str, float]] = None):
    t0 = time.perf_counter()
    try:
        yield
    finally:
        dt_s = time.perf_counter() - t0
        if store is not None:
            store[label] = dt_s
        if enabled:
            print(f"[timing] {label}={dt_s:.2f}s", flush=True)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker", default=None, help="Ticker (required unless --cik is provided)")
    ap.add_argument("--cik", default=None, help="CIK as int; overrides ticker lookup")
    ap.add_argument("--out", default=None, help="Output Excel path (default: Excel stock models/{TICKER}_model.xlsm)")
    ap.add_argument("--step-a-only", action="store_true", help="Run SEC ingest + raw XBRL export only")
    ap.add_argument("--max-filings", type=int, default=None, help="Max filings to download for Step A")
    ap.add_argument(
        "--forms",
        default="10-K,10-Q,8-K,DEF 14A,DEFA14A",
        help="Comma-separated SEC forms for ingest (Step A), e.g. 10-K,10-Q,8-K,DEF 14A,DEFA14A",
    )
    ap.add_argument("--no-exhibits", action="store_true", help="Disable exhibit downloads in Step A.")
    ap.add_argument("--max-file-mb", type=int, default=25, help="Skip SEC files larger than this size in MB (when size is known).")
    ap.add_argument(
        "--materialize",
        default=True,
        action=argparse.BooleanOptionalAction,
        help="Materialize canonical SEC materials under cache/materials.",
    )
    ap.add_argument("--materialize-dir", default="", help="Optional override for materials root directory.")
    ap.add_argument(
        "--attachment-mode",
        default="smart",
        choices=["smart", "all", "off"],
        help="Attachment-aware exhibit detection mode for Step A.",
    )
    ap.add_argument(
        "--verify-cache-sha256",
        action="store_true",
        help="Recompute sha256 for cache hits during Step A instead of reusing prior index values.",
    )
    ap.add_argument(
        "--materialize-method",
        default="hardlink",
        choices=["hardlink", "copy"],
        help="How Step A materializes canonical SEC materials.",
    )
    ap.add_argument(
        "--quiet-download-logs",
        default=True,
        action=argparse.BooleanOptionalAction,
        help="Suppress routine SEC download/cache logs during Step A.",
    )
    ap.add_argument("--cache-dir", default=None, help="Cache directory (default: ../{TICKER}/sec_cache)")
    ap.add_argument(
        "--market-sync",
        action="store_true",
        help="Sync cached external market reports/parsed data for the current ticker before workbook generation.",
    )
    ap.add_argument(
        "--market-refresh",
        action="store_true",
        help="Force refresh checks for cached market-report sources before parsing.",
    )
    ap.add_argument(
        "--market-reparse",
        action="store_true",
        help="Rebuild parsed market-data cache from existing raw/bootstrap sources.",
    )
    ap.add_argument(
        "--market-only",
        action="store_true",
        help="Sync/parse/export market data cache and exit without generating the workbook.",
    )
    ap.add_argument("--max-quarters", type=int, default=80, help="How many quarters back")
    ap.add_argument("--min-year", type=int, default=None, help="Drop quarters earlier than this year (e.g., 2009)")
    ap.add_argument("--price", type=float, default=None, help="Optional price to prefill Valuation inputs")
    ap.add_argument(
        "--profile-timings",
        default=True,
        action=argparse.BooleanOptionalAction,
        help="Print per-stage timing so slow stages are visible.",
    )
    ap.add_argument(
        "--quiet-pdf-warnings",
        default=True,
        action=argparse.BooleanOptionalAction,
        help="Suppress noisy pdfminer/pdfplumber warnings during PDF extraction.",
    )
    ap.add_argument(
        "--rebuild-doc-text-cache",
        action="store_true",
        help="Force re-extraction of cached PDF text under cache_dir/doc_text_cache.",
    )
    ap.add_argument(
        "--debug-regression-gate",
        action="store_true",
        help="Write/print detailed regression-gate mismatch diagnostics on failure.",
    )
    ap.add_argument(
        "--allow-regression-gate-fail",
        action="store_true",
        help="After writing regression-gate diagnostics, continue workbook generation for debugging.",
    )
    ap.add_argument("--disable-tier2", action="store_true", help="Disable Tier 2 debt tranches parsing")
    ap.add_argument("--disable-tier3", action="store_true", help="Disable Tier 3 non-GAAP parsing")
    ap.add_argument(
        "--skip-doc-intel",
        action="store_true",
        help="Require doc_intel outputs to come from stage cache instead of rebuilding them.",
    )
    ap.add_argument(
        "--only-write-excel",
        action="store_true",
        help="Reuse the latest cached pipeline bundle and only rebuild the workbook.",
    )
    ap.add_argument(
        "--rebuild-pipeline-cache",
        action="store_true",
        help="Ignore cached full pipeline bundle and rebuild it before writing Excel.",
    )
    ap.add_argument(
        "--skip-macro-injection",
        action="store_true",
        help="Write a macro-free .xlsx directly for faster development iterations.",
    )
    ap.add_argument(
        "--quarter-notes-audit",
        default=True,
        action=argparse.BooleanOptionalAction,
        help="Write visible Quarter_Notes_Audit diagnostics using saved-workbook readback as truth.",
    )
    ap.add_argument("--non-gaap-mode", default="strict", choices=["strict", "relaxed"], help="Non-GAAP parsing mode")
    ap.add_argument("--strictness", default="ytd", choices=["ytd", "only3m"], help="YTD derivation policy")
    ap.add_argument(
        "--non-gaap-preview",
        default=True,
        action=argparse.BooleanOptionalAction,
        help="When running strict mode, also generate relaxed preview sheets.",
    )
    ap.add_argument(
        "--history-export",
        default=True,
        action=argparse.BooleanOptionalAction,
        help="Write History_Q CSV/Parquet sidecars next to the Excel output.",
    )
    ap.add_argument("--history-csv", default="", help="Optional path for History_Q CSV export.")
    ap.add_argument("--history-parquet", default="", help="Optional path for History_Q Parquet export.")
    ap.add_argument(
        "--excel-mode",
        default="clean",
        choices=["clean", "full"],
        help="Excel output mode: clean (core sheets) or full (includes diagnostics + relaxed preview).",
    )
    ap.add_argument(
        "--user-agent",
        default=SEC_USER_AGENT,
        help="REQUIRED: identify yourself for SEC fair access (org/name + contact).",
    )
    args = ap.parse_args()

    ua = _require_user_agent(args.user_agent)

    tkr_u = str(args.ticker or "").upper() or "SEC"
    if args.cache_dir:
        cache_dir = Path(args.cache_dir).expanduser().resolve()
    else:
        cache_dir = canonical_ticker_cache_root(_project_root(), tkr_u).resolve()
        migration = bootstrap_canonical_ticker_cache(_project_root(), tkr_u)
        if migration.get("status") == "copied":
            print(
                "[sec_cache] "
                f"seeded canonical cache {migration['canonical']} "
                f"from legacy {migration['legacy']} "
                f"(files={migration.get('copied_files', 0)} dirs={migration.get('created_dirs', 0)}"
                f"{'; skipped=' + ','.join(migration.get('skipped', [])) if migration.get('skipped') else ''})"
            )

    min_year = args.min_year

    cfg = PipelineConfig(
        cache_dir=cache_dir,
        max_quarters=args.max_quarters,
        enable_tier2_debt=not args.disable_tier2,
        enable_tier3_non_gaap=not args.disable_tier3,
        non_gaap_mode=args.non_gaap_mode,
        strictness=args.strictness,
        non_gaap_preview=bool(args.non_gaap_preview),
        min_year=min_year,
        price=args.price,
        quiet_pdf_warnings=bool(args.quiet_pdf_warnings),
        rebuild_doc_text_cache=bool(args.rebuild_doc_text_cache),
        use_cached_doc_intel_only=bool(args.skip_doc_intel),
        profile_timings=bool(args.profile_timings),
        debug_regression_gate=bool(args.debug_regression_gate),
        allow_regression_gate_fail=bool(args.allow_regression_gate_fail),
    )

    sec_cfg = SecConfig(user_agent=ua)

    print(f"[pbi_xbrl] version {__version__}")
    print(
        "[Config] "
        f"ticker={args.ticker} "
        f"cik={args.cik or 'auto'} "
        f"cache_dir={cfg.cache_dir} "
        f"max_quarters={cfg.max_quarters} "
        f"min_year={cfg.min_year or 'None'} "
        f"tier2={'on' if cfg.enable_tier2_debt else 'off'} "
        f"tier3={'on' if cfg.enable_tier3_non_gaap else 'off'} "
        f"non_gaap_mode={cfg.non_gaap_mode} "
        f"strictness={cfg.strictness} "
        f"preview={'on' if cfg.non_gaap_preview else 'off'} "
        f"quiet_pdf={'on' if cfg.quiet_pdf_warnings else 'off'} "
        f"rebuild_doc_cache={'on' if cfg.rebuild_doc_text_cache else 'off'} "
        f"skip_doc_intel={'on' if cfg.use_cached_doc_intel_only else 'off'} "
        f"profile_timings={'on' if cfg.profile_timings else 'off'} "
        f"debug_regression_gate={'on' if cfg.debug_regression_gate else 'off'} "
        f"allow_regression_gate_fail={'on' if cfg.allow_regression_gate_fail else 'off'}"
    )

    market_requested = bool(args.market_sync or args.market_refresh or args.market_reparse or args.market_only)
    if market_requested:
        if not str(args.ticker or "").strip():
            raise SystemExit("ERROR: --ticker is required for --market-sync/--market-refresh/--market-reparse/--market-only.")
        market_profile = get_company_profile(args.ticker)
        market_timings: Dict[str, float] = {}
        with _timed("market_cache", enabled=cfg.profile_timings, store=market_timings):
            market_summary = sync_market_cache(
                cache_dir=cfg.cache_dir,
                ticker=str(args.ticker or "").upper(),
                profile=market_profile,
                sync_raw=bool(args.market_sync or args.market_refresh),
                refresh=bool(args.market_refresh),
                reparse=bool(args.market_reparse),
            )
        print(
            "[market_data] "
            f"sources={','.join(market_summary.sources_enabled) or 'none'} "
            f"raw_added={market_summary.raw_added} "
            f"raw_refreshed={market_summary.raw_refreshed} "
            f"raw_skipped={market_summary.raw_skipped} "
            f"parsed={','.join(market_summary.parsed_sources) or 'none'} "
            f"export_rows={market_summary.export_rows} "
            f"export_path={market_summary.export_path}",
            flush=True,
        )
        if args.market_only:
            return

    if args.step_a_only:
        forms = tuple(x.strip() for x in str(args.forms or "").split(",") if x.strip())
        materialize_dir = Path(args.materialize_dir).expanduser().resolve() if str(args.materialize_dir or "").strip() else None
        ingest_cfg = IngestConfig(
            cache_dir=cfg.cache_dir,
            user_agent=ua,
            forms=forms or ("10-Q", "10-K", "8-K", "DEF 14A", "DEFA14A"),
            include_exhibits=not bool(args.no_exhibits),
            max_file_mb=max(1, int(args.max_file_mb or 25)),
            materialize=bool(args.materialize),
            materialize_dir=materialize_dir,
            attachment_mode=str(args.attachment_mode or "smart").lower(),
            verify_cache_sha256=bool(args.verify_cache_sha256),
            materialize_method=str(args.materialize_method or "hardlink").lower(),
            quiet_download_logs=bool(args.quiet_download_logs),
            max_filings=args.max_filings,
        )
        filings_df, files_df, exhibits_df, instance_paths = download_all(
            ingest_cfg,
            ticker=args.ticker,
            cik=int(args.cik) if args.cik else None,
        )

        # SEC index exports
        tkr_idx = str(args.ticker or "SEC").upper()
        index_dir = (cfg.cache_dir / "sec_index")
        index_dir.mkdir(parents=True, exist_ok=True)
        filings_idx = index_dir / "filings.csv"
        files_idx = index_dir / "files.csv"
        exhibits_idx = index_dir / "exhibits.csv"
        if filings_df is not None:
            filings_df.to_csv(filings_idx, index=False)
        if files_df is not None:
            files_df.to_csv(files_idx, index=False)
        if exhibits_df is not None:
            exhibits_df.to_csv(exhibits_idx, index=False)

        # concise ingest summary
        primary_ct = int((files_df.get("kind", pd.Series(dtype=str)).astype(str).str.lower() == "primary").sum()) if isinstance(files_df, pd.DataFrame) and not files_df.empty else 0
        exhibit_ct = int((files_df.get("kind", pd.Series(dtype=str)).astype(str).str.lower() == "exhibit").sum()) if isinstance(files_df, pd.DataFrame) and not files_df.empty else 0
        skipped_size_ct = int((files_df.get("status", pd.Series(dtype=str)).astype(str).str.lower() == "skipped_size").sum()) if isinstance(files_df, pd.DataFrame) and not files_df.empty else 0
        max_row = None
        if isinstance(files_df, pd.DataFrame) and not files_df.empty and "bytes" in files_df.columns:
            bytes_num = pd.to_numeric(files_df["bytes"], errors="coerce")
            if bytes_num.notna().any():
                idx_max = int(bytes_num.idxmax())
                max_row = files_df.loc[idx_max]
        largest_txt = "n/a"
        if max_row is not None:
            b = pd.to_numeric(max_row.get("bytes"), errors="coerce")
            b_int = int(b) if pd.notna(b) else 0
            largest_txt = f"{max_row.get('filename') or max_row.get('local_path')} ({b_int:,} bytes)"
        print(
            f"[Step A Summary] ticker={tkr_idx} filings={len(filings_df) if filings_df is not None else 0} "
            f"primary={primary_ct} exhibits={exhibit_ct} skipped_size={skipped_size_ct} largest={largest_txt}"
        )
        print(f"[Step A Index] filings={filings_idx} files={files_idx} exhibits={exhibits_idx}")

        # Parse instance facts
        facts_rows = []
        for item in instance_paths:
            path = item["path"]
            try:
                data = Path(path).read_bytes()
            except Exception:
                continue
            meta = InstanceMetadata(
                accession=item.get("accession", "unknown"),
                form=item.get("form", "unknown"),
                filedDate=item.get("filedDate"),
                reportDate=item.get("reportDate"),
                primaryDoc=item.get("primaryDoc"),
            )
            df = parse_instance(data, meta)
            if not df.empty:
                df["instance_path"] = str(path)
                facts_rows.append(df)
        facts_df = pd.concat(facts_rows, ignore_index=True) if facts_rows else pd.DataFrame()

        default_step_a_out = _default_step_a_out_path(args.ticker)
        out_path = default_step_a_out if not args.out else _normalize_out_path_xlsx(args.out)
        if out_path == _default_out_path(args.ticker):
            out_path = default_step_a_out
        wb = Workbook()
        wb.remove(wb.active)

        def _write_sheet(name: str, df: pd.DataFrame) -> None:
            ws = wb.create_sheet(title=name)
            if df is None or df.empty:
                ws["A1"] = "No data."
                return
            ws.append(list(df.columns))
            for _, row in df.iterrows():
                ws.append([row.get(c) for c in df.columns])
            # table style
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName=f"{name}Tbl", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
            ws.freeze_panes = "A2"
            # autofit
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        _write_sheet("Filings", filings_df)
        _write_sheet("DownloadManifest", files_df)
        _write_sheet("XBRL_Facts_Raw", facts_df)
        _write_sheet("Exhibits_Index", exhibits_df)

        # Errors/Warnings summary
        errors = []
        if files_df is not None and not files_df.empty:
            err_counts = files_df["status"].value_counts().to_dict()
            for k, v in err_counts.items():
                errors.append({"type": k, "count": v})
        _write_sheet("Errors_Warnings", pd.DataFrame(errors))

        wb.save(out_path)
        print(f"[OK] Wrote Excel: {out_path}")
        return

    repo_root = Path(__file__).resolve().parents[1]
    pipeline_cache_key = _pipeline_bundle_cache_key(args, cfg, repo_root)
    pipeline_bundle = None
    timing_rows: Dict[str, float] = {}
    if args.only_write_excel:
        pipeline_bundle = _load_pipeline_bundle_cache(cfg.cache_dir, args.ticker, pipeline_cache_key, ignore_key=True)
        if pipeline_bundle is None:
            raise SystemExit("ERROR: --only-write-excel requires an existing cached pipeline bundle.")
    elif not args.rebuild_pipeline_cache:
        pipeline_bundle = _load_pipeline_bundle_cache(cfg.cache_dir, args.ticker, pipeline_cache_key, ignore_key=False)

    if pipeline_bundle is None:
        with _timed("run_pipeline", enabled=cfg.profile_timings, store=timing_rows):
            pipeline_bundle = run_pipeline(
                cfg,
                sec_cfg,
                ticker=args.ticker,
                cik=args.cik,
            )
        _save_pipeline_bundle_cache(cfg.cache_dir, args.ticker, pipeline_cache_key, pipeline_bundle)

    (
        hist,
        audit,
        debt_tranches,
        debt_recon,
        adj_metrics,
        adj_breakdown,
        non_gaap_files,
        adj_metrics_relaxed,
        adj_breakdown_relaxed,
        non_gaap_files_relaxed,
        needs_review,
        info_log,
        tag_coverage,
        period_checks,
        qa_checks,
        bridge_q,
        manifest_df,
        ocr_log,
        qfd_preview,
        qfd_unused,
        debt_profile,
        debt_tranches_latest,
        debt_maturity,
        debt_credit_notes,
        revolver_df,
        revolver_history,
        debt_buckets,
        slides_segments,
        slides_debt,
        slides_guidance,
        quarter_notes,
        promises,
        promise_progress,
        non_gaap_cred,
        company_overview,
    ) = pipeline_bundle

    out_path = _default_out_path(args.ticker) if not args.out else _normalize_out_path_xlsm(args.out)
    xlsx_tmp_path = out_path.with_name(f"{out_path.stem}_nomacro.xlsx")
    with _timed("write_excel", enabled=cfg.profile_timings, store=timing_rows):
        writer_result = write_excel(
            out_path=xlsx_tmp_path,
            hist=hist,
            audit=audit,
            needs_review=needs_review,
            debt_tranches=debt_tranches,
            debt_recon=debt_recon,
            adj_metrics=adj_metrics,
            adj_breakdown=adj_breakdown,
            non_gaap_files=non_gaap_files,
            adj_metrics_relaxed=adj_metrics_relaxed,
            adj_breakdown_relaxed=adj_breakdown_relaxed,
            non_gaap_files_relaxed=non_gaap_files_relaxed,
            info_log=info_log,
            tag_coverage=tag_coverage,
            period_checks=period_checks,
            qa_checks=qa_checks,
            bridge_q=bridge_q,
            manifest_df=manifest_df,
            ocr_log=ocr_log,
            qfd_preview=qfd_preview,
            qfd_unused=qfd_unused,
            debt_profile=debt_profile,
            debt_tranches_latest=debt_tranches_latest,
            debt_maturity=debt_maturity,
            debt_credit_notes=debt_credit_notes,
            revolver_df=revolver_df,
            revolver_history=revolver_history,
            debt_buckets=debt_buckets,
            slides_segments=slides_segments,
            slides_debt=slides_debt,
            slides_guidance=slides_guidance,
            quarter_notes=quarter_notes,
            promises=promises,
            promise_progress=promise_progress,
            non_gaap_cred=non_gaap_cred,
            company_overview=company_overview,
            ticker=args.ticker,
            price=cfg.price,
            strictness=cfg.strictness,
            excel_mode=args.excel_mode,
            is_rules=get_income_statement_rules(args.ticker),
            cache_dir=cfg.cache_dir,
            quiet_pdf_warnings=cfg.quiet_pdf_warnings,
            rebuild_doc_text_cache=cfg.rebuild_doc_text_cache,
            profile_timings=cfg.profile_timings,
            quarter_notes_audit=bool(args.quarter_notes_audit),
            capture_saved_workbook_provenance=False,
        )
    try:
        if out_path.exists():
            out_path.unlink()
    except Exception:
        pass
    if args.skip_macro_injection:
        final_out_path = out_path.with_suffix(".xlsx")
        try:
            if final_out_path.exists():
                final_out_path.unlink()
        except Exception:
            pass
        try:
            xlsx_tmp_path.replace(final_out_path)
        except Exception:
            final_out_path = xlsx_tmp_path
        timing_rows["macro_injection"] = 0.0
        print("[Info] Macro injection skipped by flag.", flush=True)
    else:
        try:
            with _timed("macro_injection", enabled=cfg.profile_timings, store=timing_rows):
                inject_valuation_macros(
                    xlsx_tmp_path,
                    out_path,
                    worksheet_name="Valuation",
                    debug_log_path=cfg.cache_dir / "xlsm_injection_debug.log",
                )
            final_out_path = out_path
        except Exception as e:
            final_out_path = out_path.with_suffix(".xlsx")
            try:
                if final_out_path.exists():
                    final_out_path.unlink()
            except Exception:
                pass
            try:
                xlsx_tmp_path.replace(final_out_path)
            except Exception:
                final_out_path = xlsx_tmp_path
            reason_hint = ""
            if isinstance(e, MacroInjectionError) and e.failed_step == "vbproject_access":
                reason_hint = (
                    "Verify Excel Trust Center setting 'Trust access to the VBA project object model' "
                    "if VBA project access is still blocked. "
                )
            failed_step_text = ""
            if isinstance(e, MacroInjectionError) and e.failed_step:
                failed_step_text = f" Failed step: {e.failed_step}."
            debug_log_text = ""
            if isinstance(e, MacroInjectionError) and e.debug_log_path is not None:
                debug_log_text = f" Debug log: {e.debug_log_path}."
            print(
                "WARN: VBA injection skipped. "
                f"{reason_hint}"
                f"Wrote macro-free workbook instead: {final_out_path}. "
                f"Details: {type(e).__name__}: {e}."
                f"{failed_step_text}"
                f"{debug_log_text}"
            )
        finally:
            try:
                if xlsx_tmp_path.exists():
                    xlsx_tmp_path.unlink()
            except Exception:
                pass

    final_provenance = _verify_saved_workbook_export(final_out_path, writer_result)
    if bool(args.quarter_notes_audit) and getattr(writer_result, "quarter_notes_audit_rows", None):
        final_audit_rows = enrich_quarter_notes_audit_rows_with_readback(
            list(getattr(writer_result, "quarter_notes_audit_rows", []) or []),
            final_provenance,
        )
        write_quarter_notes_audit_sheet(final_out_path, final_audit_rows)
        try:
            validate_saved_workbook_after_audit_write(
                final_out_path,
                quarter_notes_ui_snapshot=getattr(writer_result, "quarter_notes_ui_snapshot", {}) or {},
                quarter_notes_header_text=str(getattr(writer_result, "quarter_notes_header_text", "") or ""),
            )
        except Exception as exc:
            raise SystemExit(
                f"ERROR: saved workbook audit verification failed for {final_out_path}: {type(exc).__name__}: {exc}"
            ) from exc

    print(f"[OK] Wrote Excel: {final_out_path}")
    print(
        f"[Info] quarters={len(hist)}; "
        f"tier2_debt_rows={len(debt_tranches)}; "
        f"tier3_adj_quarters={len(adj_metrics)}; "
        f"needs_review={len(needs_review)}"
    )

    if args.history_export:
        csv_path = Path(args.history_csv).expanduser().resolve() if args.history_csv else _default_history_export_path(args.ticker, ".csv")
        with _timed("history_csv_export", enabled=cfg.profile_timings, store=timing_rows):
            hist.to_csv(csv_path, index=False)
        print(f"[OK] Wrote CSV: {csv_path}")
        parquet_path = Path(args.history_parquet).expanduser().resolve() if args.history_parquet else _default_history_export_path(args.ticker, ".parquet")
        try:
            with _timed("history_parquet_export", enabled=cfg.profile_timings, store=timing_rows):
                hist.to_parquet(parquet_path, index=False)
            print(f"[OK] Wrote Parquet: {parquet_path}")
        except Exception as e:
            print(f"[WARN] Parquet export skipped: {type(e).__name__}: {e}")

    if cfg.profile_timings and timing_rows:
        summary = " | ".join(f"{k}={v:.2f}s" for k, v in sorted(timing_rows.items(), key=lambda kv: (-kv[1], kv[0])))
        print(f"[Timing Summary] {summary}", flush=True)


if __name__ == "__main__":
    main()
