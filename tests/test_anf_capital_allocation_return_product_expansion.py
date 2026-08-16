from __future__ import annotations

from copy import deepcopy
import json
from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
import pytest

from pbi_xbrl.json_schema_validation import load_json_strict
from pbi_xbrl.longitudinal_memory.capital_allocation_return_product_expansion import (
    CapitalAllocationReturnExpansionError,
    EXPECTED_ACCEPTED_PREVIEW_SHA256,
    build_capital_allocation_return_investor_product,
    build_capital_allocation_return_workbook_projection_plan,
    materialize_capital_allocation_return_workbook_projection,
)
from pbi_xbrl.longitudinal_memory.summary_bs_workbook_materialization import (
    _sheet_part_map,
    sha256_file,
)


ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = Path(r"C:\Users\Jibbe\Aktier\StockModelData")
PACKAGE = (
    DATA_ROOT
    / "outputs"
    / "stress_tests"
    / "ANF_new_ticker_engine"
    / "ANF_normalized_data_package.json"
)
BASE = (
    DATA_ROOT
    / "audit"
    / "capital_return_debt_bounded_correction_2026-08-16"
    / "ANF_capital_return_debt_source_native_preview_a.xlsx"
)
BS_PRODUCT = ROOT / "tests" / "fixtures" / "summary_bs" / "anf_bs_segment_product.v1.json"
BS_SHADOW = ROOT / "tests" / "fixtures" / "summary_bs" / "anf_bs_segment_shadow.v1.json"
CALC_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _require_inputs() -> None:
    missing = [str(path) for path in (PACKAGE, BASE, BS_PRODUCT, BS_SHADOW) if not path.exists()]
    if missing:
        pytest.skip(f"Accepted local product inputs are unavailable: {missing!r}")


@pytest.fixture(scope="session")
def package() -> dict:
    _require_inputs()
    return load_json_strict(PACKAGE)


@pytest.fixture(scope="session")
def bs_product() -> dict:
    _require_inputs()
    return load_json_strict(BS_PRODUCT)


@pytest.fixture(scope="session")
def bs_shadow() -> dict:
    _require_inputs()
    return load_json_strict(BS_SHADOW)


@pytest.fixture(scope="session")
def product(package: dict, bs_product: dict, bs_shadow: dict):
    return build_capital_allocation_return_investor_product(
        package=package,
        balance_sheet_product=bs_product,
        balance_sheet_shadow=bs_shadow,
    )


@pytest.fixture(scope="session")
def plan(package: dict, bs_product: dict, bs_shadow: dict):
    return build_capital_allocation_return_workbook_projection_plan(
        package=package,
        source_package_path=PACKAGE,
        balance_sheet_product=bs_product,
        balance_sheet_product_path=BS_PRODUCT,
        balance_sheet_shadow=bs_shadow,
        balance_sheet_shadow_path=BS_SHADOW,
        base_workbook=BASE,
    )


def _rows(rows) -> dict[str, dict]:
    return {str(row["row_key"]): row for row in rows}


def _formula_map(workbook) -> dict[tuple[str, str], str]:
    return {
        (sheet.title, cell.coordinate): cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }


def _defined_names(path: Path) -> dict[tuple[str, str | None], str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    names = root.find("m:definedNames", CALC_NS)
    return {
        (row.attrib["name"], row.attrib.get("localSheetId")): row.text or ""
        for row in (() if names is None else names)
    }


def _calc_properties(path: Path) -> dict[str, str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    calc = root.find("m:calcPr", CALC_NS)
    assert calc is not None
    return dict(calc.attrib)


def test_product_sections_owners_and_relevance(product) -> None:
    assert product.summary_periods == ("2026-Q1", "TTM through 2026-Q1", "2025-FY")
    assert product.annual_allocation_periods == (
        "2021-FY",
        "2022-FY",
        "2023-FY",
        "2024-FY",
        "2025-FY",
    )
    assert product.quarterly_return_periods == (
        "2023-Q2",
        "2023-Q3",
        "2023-Q4",
        "2024-Q1",
        "2024-Q2",
        "2024-Q3",
        "2024-Q4",
        "2025-Q1",
        "2025-Q2",
        "2025-Q3",
        "2025-Q4",
        "2026-Q1",
    )
    assert product.annual_return_periods == ("2024-FY", "2025-FY")
    assert tuple(len(getattr(product, key)) for key in (
        "capital_allocation_summary",
        "annual_capital_allocation_history",
        "capital_return_summary",
        "quarterly_capital_return_history",
        "annual_capital_return_history",
    )) == (4, 4, 8, 6, 6)
    assert all(row["owner"] != "capital_allocation" for row in product.capital_allocation_owner_map)
    assert sum(row["displayed"] for row in product.capital_allocation_owner_map) == 4


def test_summary_economics_and_point_in_time_rules(product) -> None:
    allocation = _rows(product.capital_allocation_summary)
    assert [value["value"] for value in allocation["free_cash_flow"]["values"]] == pytest.approx(
        [-17.085, 416.047, 378.368]
    )
    assert [value["value"] for value in allocation["capital_expenditures"]["values"]] == pytest.approx(
        [61.341, 251.351, 240.774]
    )
    assert [value["value"] for value in allocation["repurchase_cash_program"]["values"]] == pytest.approx(
        [105.018, 356.242, 451.224]
    )
    net_cash = allocation["ending_net_cash"]["values"]
    assert [value["value"] for value in net_cash] == pytest.approx([619.224, 619.224, 784.576])
    assert net_cash[1]["period_behavior"] == "terminal_point_in_time"

    returns = _rows(product.capital_return_summary)
    assert returns["share_issuance_sbc"]["values"][0]["value"] == pytest.approx(0.582)
    assert returns["net_share_reduction"]["values"][0]["value"] == pytest.approx(0.574)
    assert returns["buybacks_to_fcf"]["values"][0]["value"] is None
    assert all(value["value"] is None for value in returns["dividends_paid"]["values"])
    authorization = returns["authorization_remaining"]["values"]
    assert authorization[0]["value"] == pytest.approx(745.080737)
    assert authorization[1]["value"] == pytest.approx(745.080737)


def test_history_coverage_issuance_and_annual_derivations(product) -> None:
    allocation = _rows(product.annual_capital_allocation_history)
    assert sum(value["status"] == "available" for row in allocation.values() for value in row["values"]) == 14
    assert [value["value"] for value in allocation["repurchase_cash_program"]["values"]] == [
        None,
        None,
        None,
        pytest.approx(229.807),
        pytest.approx(451.224),
    ]
    quarterly = _rows(product.quarterly_capital_return_history)
    assert sum(value["status"] == "available" for row in quarterly.values() for value in row["values"]) == 52
    issuance = quarterly["share_issuance_sbc"]["values"]
    assert [value["value"] for value in issuance[:3]] == [None, None, None]
    assert [value["value"] for value in issuance[-3:]] == pytest.approx([0.039, 0.016, 0.582])
    annual = _rows(product.annual_capital_return_history)
    assert sum(value["status"] == "available" for row in annual.values() for value in row["values"]) == 12
    checks = product.derivation_review["annual_average_price"]
    assert all(row["matches_accepted_derivation"] for row in checks)
    assert all(row["simple_average_differs"] for row in checks)
    assert product.derivation_review["simple_average_of_quarterly_ratio_count"] == 0


def test_mutation_guards_and_new_ticker_fail_closed(
    package: dict, bs_product: dict, bs_shadow: dict, product
) -> None:
    forward = deepcopy(package)
    forward.setdefault("investment_case", {})["forward_buyback_assumption"] = 999999.0
    assert build_capital_allocation_return_investor_product(
        package=forward,
        balance_sheet_product=bs_product,
        balance_sheet_shadow=bs_shadow,
    ).product_digest == product.product_digest

    debt = deepcopy(package)
    debt.setdefault("debt_liquidity", {})["revolver_availability"] = 999999.0
    assert build_capital_allocation_return_investor_product(
        package=debt,
        balance_sheet_product=bs_product,
        balance_sheet_shadow=bs_shadow,
    ).product_digest == product.product_digest

    missing_issuance = deepcopy(package)
    missing_issuance["capital_returns"]["records"] = [
        row
        for row in missing_issuance["capital_returns"]["records"]
        if not (
            row["metric_id"] == "share_issuance_sbc"
            and row["fiscal_period"] == "2026-Q1"
            and row["period_type"] == "quarter"
        )
    ]
    missing_product = build_capital_allocation_return_investor_product(
        package=missing_issuance,
        balance_sheet_product=bs_product,
        balance_sheet_shadow=bs_shadow,
    )
    assert _rows(missing_product.capital_return_summary)["share_issuance_sbc"]["values"][0]["value"] is None

    bad_average = deepcopy(package)
    check = product.derivation_review["annual_average_price"][0]
    for row in bad_average["capital_returns"]["records"]:
        if row["metric_id"] == "cash_per_program_share" and row["fiscal_period"] == "2024-FY":
            row["value"] = check["quarterly_simple_average"]
    with pytest.raises(CapitalAllocationReturnExpansionError, match="repurchase-price"):
        build_capital_allocation_return_investor_product(
            package=bad_average,
            balance_sheet_product=bs_product,
            balance_sheet_shadow=bs_shadow,
        )

    bad_authorization = deepcopy(package)
    for row in bad_authorization["capital_returns"]["records"]:
        if row["metric_id"] == "authorization_remaining" and row["period_type"] == "ttm":
            row["value"] = 999999.0
    with pytest.raises(CapitalAllocationReturnExpansionError, match="terminal"):
        build_capital_allocation_return_investor_product(
            package=bad_authorization,
            balance_sheet_product=bs_product,
            balance_sheet_shadow=bs_shadow,
        )

    no_returns = deepcopy(package)
    no_returns.pop("capital_returns", None)
    no_return_product = build_capital_allocation_return_investor_product(
        package=no_returns,
        balance_sheet_product=bs_product,
        balance_sheet_shadow=bs_shadow,
    )
    no_return_allocation = _rows(no_return_product.capital_allocation_summary)
    assert set(no_return_allocation) == {
        "free_cash_flow",
        "capital_expenditures",
        "ending_net_cash",
    }
    assert [
        value["value"] for value in no_return_allocation["free_cash_flow"]["values"]
    ] == pytest.approx([-17.085, 416.047, 378.368])
    assert "repurchase_cash_program" not in no_return_allocation
    assert no_return_product.capital_return_summary == ()
    assert no_return_product.quarterly_capital_return_history == ()
    assert no_return_product.annual_capital_return_history == ()


def test_binding_layout_and_current_45_slot_disposition(plan) -> None:
    assert plan.base_workbook_sha256 == EXPECTED_ACCEPTED_PREVIEW_SHA256
    assert sha256_file(BASE) == EXPECTED_ACCEPTED_PREVIEW_SHA256
    assert len(plan.bindings) == 140
    assert sum(row["status"] == "available" for row in plan.bindings) == 110
    assert len({row["target_cell"] for row in plan.bindings}) == 140
    assert all(not row["target_cell"].startswith("Valuation!A19") for row in plan.bindings)
    dispositions = plan.investor_product["current_45_slot_disposition"]
    assert len(dispositions) == 45
    assert {row["disposition"] for row in dispositions} == {
        "MOVED_TO_SUMMARY",
        "HIDDEN_SUPPORT_ONLY",
        "INTENTIONALLY_NOT_DISPLAYED",
        "UNAVAILABLE",
    }
    assert sum(row["disposition"] == "MOVED_TO_SUMMARY" for row in dispositions) == 24
    assert all(row.mode != "SET_FORMULA" for row in plan.cell_mutations)
    assert not plan.bindings[0]["owner"].startswith("capital_allocation")


def test_materialized_readback_lineage_and_visual_contract(tmp_path: Path, plan) -> None:
    output = tmp_path / "expanded.xlsx"
    result = materialize_capital_allocation_return_workbook_projection(
        plan=plan,
        base_workbook=BASE,
        output_workbook=output,
    )
    workbook = load_workbook(output, data_only=False)
    try:
        valuation = workbook["Valuation"]
        assert valuation["N79"].value == "Capital Allocation & Capital Return"
        assert valuation["N81"].value == "A. Capital Allocation Summary"
        assert valuation["N88"].value == "B. Annual Capital Allocation History"
        assert valuation["N95"].value == "C. Capital Return Summary"
        assert valuation["N106"].value == "D. Quarterly Capital Return History"
        assert valuation["N115"].value == "E. Annual Capital Return History"
        assert [valuation.cell(107, column).value for column in range(16, 28)] == [
            "Q2'23", "Q3'23", "Q4'23", "Q1'24", "Q2'24", "Q3'24",
            "Q4'24", "Q1'25", "Q2'25", "Q3'25", "Q4'25", "Q1'26",
        ]
        assert valuation["P83"].value == pytest.approx(-17.085)
        assert valuation["Q83"].value == pytest.approx(416.047)
        assert valuation["R83"].value == pytest.approx(378.368)
        assert valuation["P97"].value == pytest.approx(105.018)
        assert valuation["P100"].value == pytest.approx(0.582)
        assert valuation["P101"].value == pytest.approx(0.574)
        assert valuation["P102"].value is None
        assert valuation["P103"].value is None
        assert valuation["Q104"].value == pytest.approx(745.080737)
        assert valuation["A151"].value == "Capital allocation & return"
        assert valuation["A152"].value == (
            "See Capital Allocation & Capital Return beside the historical Valuation grid."
        )
        assert all(valuation.row_dimensions[row].hidden for row in range(153, 169))
        assert valuation["B194"].value == "=IC_Current_GAAP_EPS"
        assert valuation["E198"].value == "=IC_Bull_Upside_Downside"
        assert valuation["AI139"].value == '=IFERROR(MATCH(1,\'Hidden_Value_Flags\'!$L$2:$L$100,0)+1,"")'
        support_cells = [
            valuation.cell(row=row, column=column).value
            for row in range(153, 159)
            for column in range(1, 6)
            if valuation.cell(row=row, column=column).value
        ]
        assert len(support_cells) == 28
        assert all(valuation.row_dimensions[row].hidden for row in range(153, 159))
        for raw_support in support_cells:
            support = json.loads(raw_support)
            assert support["bindings"]
            assert support["metric_id"]
            assert support["support_digest"]
        assert "N79:AA79" in {str(item) for item in valuation.merged_cells.ranges}
        assert "N107:O107" in {str(item) for item in valuation.merged_cells.ranges}
        assert valuation["P108"].number_format == "#,##0.0"
        assert valuation["P110"].number_format == "$0.00"
        assert valuation["P113"].number_format == "0.0%"
    finally:
        workbook.close()
    assert result.cell_mutation_count == len(plan.cell_mutations)
    assert result.write_type_counts.get("formula", 0) == 0
    assert set(result.changed_ooxml_parts) == {"xl/styles.xml", "xl/worksheets/sheet2.xml"}


def test_lossless_preservation_and_formula_ownership(tmp_path: Path, plan) -> None:
    output = tmp_path / "expanded.xlsx"
    materialize_capital_allocation_return_workbook_projection(
        plan=plan,
        base_workbook=BASE,
        output_workbook=output,
    )
    base_workbook = load_workbook(BASE, data_only=False)
    output_workbook = load_workbook(output, data_only=False)
    try:
        assert _formula_map(output_workbook) == _formula_map(base_workbook)
        assert len(_formula_map(output_workbook)) == len(_formula_map(base_workbook))
        assert sum(
            1
            for (sheet, cell) in _formula_map(output_workbook)
            if sheet == "Valuation" and cell != "AI139"
        ) == 20
        assert base_workbook["Valuation"].freeze_panes == output_workbook["Valuation"].freeze_panes
        assert base_workbook["Valuation"].sheet_view.zoomScale == output_workbook["Valuation"].sheet_view.zoomScale
        for sheet in base_workbook.sheetnames:
            assert base_workbook[sheet].sheet_state == output_workbook[sheet].sheet_state
    finally:
        base_workbook.close()
        output_workbook.close()
    assert _defined_names(output) == _defined_names(BASE)
    assert _calc_properties(output) == _calc_properties(BASE)
    with ZipFile(BASE, "r") as before, ZipFile(output, "r") as after:
        before_parts = set(before.namelist())
        after_parts = set(after.namelist())
        assert before_parts == after_parts
        valuation_part = _sheet_part_map(before)["Valuation"]
        assert valuation_part == "xl/worksheets/sheet2.xml"
        assert all(
            before.read(part) == after.read(part)
            for part in before_parts - {valuation_part, "xl/styles.xml"}
        )


def test_all_available_bindings_are_typed_and_traceable(plan) -> None:
    available = [row for row in plan.bindings if row["status"] == "available"]
    assert len(available) == 110
    assert all(row["source_identity"] for row in available)
    assert all(row["source_ref"] for row in available)
    assert all(row["owner"] for row in available)
    assert all(row["period"] and row["source_period"] for row in available)
    assert all(row["unit"] for row in available)
    assert sum(row["value"] == 0 for row in available) >= 0
