from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
import pytest

from pbi_xbrl.json_schema_validation import load_json_strict
from pbi_xbrl.longitudinal_memory.capital_return_debt_workbook_projection import (
    EXPECTED_CAPITAL_RETURN_PROJECTION_DIGEST,
    EXPECTED_DEBT_PROJECTION_DIGEST,
    EXPECTED_VALUATION_GOLDEN_SHA256,
    build_capital_return_debt_workbook_projection_plan,
    materialize_capital_return_debt_workbook_projection,
)
from pbi_xbrl.longitudinal_memory.summary_bs_workbook_materialization import (
    canonical_ooxml_sha256,
    sha256_file,
)
from pbi_xbrl.new_ticker_investment_case_formula_surface import (
    canonical_investment_case_defined_names,
)


DATA_ROOT = Path(r"C:\Users\Jibbe\Aktier\StockModelData")
PACKAGE = (
    DATA_ROOT
    / "outputs"
    / "stress_tests"
    / "ANF_new_ticker_engine"
    / "ANF_normalized_data_package.json"
)
BASE = (
    DATA_ROOT
    / "audit"
    / "valuation_golden_acceptance_2026-08-15"
    / "golden"
    / "ANF_valuation_source_native_golden_v1.xlsx"
)
EXPECTED_BINDING_PLAN_DIGEST = (
    "b264924b949844dd93557bd1d4dd5f0fa857669abeaea19411c82d73023a194d"
)
CALCULATION_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _require_inputs() -> None:
    missing = [str(path) for path in (PACKAGE, BASE) if not path.exists()]
    if missing:
        pytest.skip(f"Local accepted Capital Return/Debt inputs are unavailable: {missing!r}")


@pytest.fixture(scope="session")
def package() -> dict:
    _require_inputs()
    return load_json_strict(PACKAGE)


@pytest.fixture(scope="session")
def plan(package: dict):
    return build_capital_return_debt_workbook_projection_plan(
        package=package,
        source_package_path=PACKAGE,
        base_workbook=BASE,
    )


def _by_target(plan) -> dict[tuple[str, str], object]:
    return {(row.target_sheet, row.target_cell): row for row in plan.cell_mutations}


def _formula_count(workbook, sheet: str) -> int:
    return sum(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for row in workbook[sheet].iter_rows()
        for cell in row
    )


def _calc_properties(path: Path) -> dict[str, str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    calc = root.find("m:calcPr", CALCULATION_NS)
    assert calc is not None
    return dict(calc.attrib)


def _defined_names(path: Path) -> dict[tuple[str, str | None], str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    names = root.find("m:definedNames", CALCULATION_NS)
    return {
        (row.attrib["name"], row.attrib.get("localSheetId")): row.text or ""
        for row in (() if names is None else names)
    }


def test_projection_plan_is_exact_and_reproducible(package: dict, plan) -> None:
    repeat = build_capital_return_debt_workbook_projection_plan(
        package=package,
        source_package_path=PACKAGE,
        base_workbook=BASE,
    )
    assert plan.to_dict() == repeat.to_dict()
    assert plan.base_workbook_sha256 == EXPECTED_VALUATION_GOLDEN_SHA256
    assert sha256_file(BASE) == EXPECTED_VALUATION_GOLDEN_SHA256
    assert plan.binding_plan_digest == EXPECTED_BINDING_PLAN_DIGEST
    assert plan.capital_return_projection["projection_digest"] == (
        EXPECTED_CAPITAL_RETURN_PROJECTION_DIGEST
    )
    assert plan.debt_projection["projection_digest"] == EXPECTED_DEBT_PROJECTION_DIGEST
    slots = [
        row[key]
        for row in plan.capital_return_projection["product_rows"]
        for key in ("latest_quarter", "ttm", "latest_completed_year")
    ]
    assert len(slots) == 45
    assert sum(value is not None for value in slots) == 27
    assert sum(value is None for value in slots) == 18
    assert len(plan.table_mutations) == 4
    assert all(row.trim_empty_tail for row in plan.dimension_mutations)


def test_stale_buyback_and_definition_mixing_mutation_guards(plan) -> None:
    targets = _by_target(plan)
    m63 = targets[("Valuation", "M63")]
    assert m63.mode == "SET_VALUE"
    assert float(m63.value) == pytest.approx(356.242)
    assert m63.semantic_owner == (
        "capital_return.repurchase_cash_program.ttm.current_consumer"
    )
    assert targets[("Valuation", "B152")].mode == "CLEAR_CONTENTS"
    rows = {row["row_key"]: row for row in plan.capital_return_projection["product_rows"]}
    assert rows["accounting_program_shares_repurchased"]["latest_quarter"] == pytest.approx(1.156)
    assert rows["cash_per_program_share"]["latest_quarter"] == pytest.approx(90.846021)
    assert rows["reported_average_all_purchases"]["latest_quarter"] == pytest.approx(90.18)
    assert rows["authorization_remaining"]["latest_quarter"] == pytest.approx(745.080737)
    assert rows["repurchase_cash_program"]["latest_quarter"] == pytest.approx(105.018)
    assert all(
        token not in mutation.semantic_owner.casefold()
        for mutation in plan.cell_mutations
        if mutation.semantic_owner.startswith("capital_return")
        for token in (
            "forward_buyback",
            "future_execution",
            "future_retired",
            "future_issuance",
            "forward_debt",
            "forward_financing",
        )
    )


def test_debt_zero_same_date_and_interest_retirement_guards(plan) -> None:
    debt_profile = plan.debt_projection["debt_profile_rows"]
    core = [row for row in debt_profile if row["category"] == "core_debt"]
    assert len(core) == 1
    assert core[0]["value"] == pytest.approx(0.0)
    assert core[0]["state"] == "reported_zero"
    assert core[0]["as_of_date"] == "2026-05-02"
    current = plan.debt_projection["leverage_liquidity_rows"][-1]
    assert current["period"] == "2026-Q1"
    assert current["as_of_date"] == "2026-05-02"
    assert current["core_debt"] == pytest.approx(0.0)
    assert current["cash"] == pytest.approx(594.080)
    assert current["revolver_availability"] == pytest.approx(449.531)
    assert current["disposition_state"] == "source_backed_reported_zero"
    assert set(current["formula_ids"]) == {
        "debt_product_net_debt",
        "debt_product_same_date_liquidity",
        "debt_product_gross_leverage",
        "debt_product_net_leverage",
    }
    targets = _by_target(plan)
    assert all(
        targets[("Leverage_Liquidity", f"K{row}")].mode == "CLEAR_CONTENTS"
        for row in range(4, 16)
    )


def test_full_projection_readback_and_formula_ownership(tmp_path: Path, plan) -> None:
    output = tmp_path / "preview.xlsx"
    result = materialize_capital_return_debt_workbook_projection(
        plan=plan,
        base_workbook=BASE,
        output_workbook=output,
    )
    workbook = load_workbook(output, data_only=False)
    try:
        valuation = workbook["Valuation"]
        slots = [valuation.cell(row, column).value for row in range(154, 169) for column in range(2, 5)]
        assert len(slots) == 45
        assert sum(value is not None for value in slots) == 27
        assert sum(value is None for value in slots) == 18
        assert valuation["M63"].value == pytest.approx(356.242)
        assert valuation["B152"].value is None
        assert (valuation["B153"].value, valuation["C153"].value, valuation["D153"].value) == (
            "Q1'26",
            "TTM Q1'26",
            "FY25",
        )
        assert valuation["A156"].value == "Cash/program share ($/share)"
        assert valuation["A162"].value == "Reported average all purchases ($/share)"
        assert _formula_count(workbook, "Valuation") == 21
        assert _formula_count(workbook, "Revolver_History") == 12
        assert _formula_count(workbook, "Leverage_Liquidity") == 48
        assert all(workbook["Leverage_Liquidity"][f"K{row}"].value is None for row in range(4, 16))
        assert workbook["Debt_Profile"]["D14"].value == pytest.approx(0.0)
        assert workbook["Debt_Profile"]["H14"].value == "reported_zero"
        assert workbook["Leverage_Liquidity"]["D15"].value == pytest.approx(0.0)
        assert workbook["Leverage_Liquidity"]["F15"].value.startswith("=IFERROR")
        assert workbook["Leverage_Liquidity"]["H15"].value.startswith("=IFERROR")
        assert workbook["Debt_Maturity_Ladder"].sheet_state == "hidden"
        assert workbook["Debt_Tranches_Latest"].sheet_state == "hidden"
        assert workbook["Debt_Tranches_Q"].sheet_state == "hidden"
        assert workbook["Debt_Buckets"].sheet_state == "hidden"
        assert workbook["Debt_Recon"].sheet_state == "hidden"
        assert workbook["Debt_Profile"].tables["Debt_Profile"].ref == "A3:J14"
        assert workbook["Revolver_History"].tables["Revolver_History"].ref == "A3:P15"
        assert workbook["Leverage_Liquidity"].tables["Leverage_Liquidity"].ref == "A3:N15"
        assert workbook["Debt_Credit_Notes"].tables["Debt_Credit_Notes"].ref == "A3:H9"
        assert all(
            workbook[sheet].tables[name].tableStyleInfo.showRowStripes is False
            for sheet, name in (
                ("Debt_Profile", "Debt_Profile"),
                ("Revolver_History", "Revolver_History"),
                ("Leverage_Liquidity", "Leverage_Liquidity"),
                ("Debt_Credit_Notes", "Debt_Credit_Notes"),
            )
        )
        assert workbook["Valuation"]["AI139"].value == (
            '=IFERROR(MATCH(1,\'Hidden_Value_Flags\'!$L$2:$L$100,0)+1,"")'
        )
        assert all(
            workbook.defined_names.get(name) is not None
            for name in canonical_investment_case_defined_names()
        )
    finally:
        workbook.close()
    assert result.table_mutation_count == 4
    assert result.write_type_counts["formula"] == 60
    assert result.write_type_counts["remove"] == 1577
    assert result.calculation_metadata_change_count == 0
    assert _calc_properties(output) == _calc_properties(BASE)


def test_lossless_scope_and_product_ownership_guards(tmp_path: Path, plan) -> None:
    output = tmp_path / "preview.xlsx"
    result = materialize_capital_return_debt_workbook_projection(
        plan=plan,
        base_workbook=BASE,
        output_workbook=output,
    )
    expected_changed = {
        "xl/styles.xml",
        "xl/tables/table2.xml",
        "xl/tables/table3.xml",
        "xl/tables/table4.xml",
        "xl/tables/table5.xml",
        "xl/workbook.xml",
        "xl/worksheets/sheet2.xml",
        "xl/worksheets/sheet9.xml",
        "xl/worksheets/sheet11.xml",
        "xl/worksheets/sheet12.xml",
        "xl/worksheets/sheet16.xml",
        "xl/worksheets/sheet17.xml",
    }
    assert set(result.changed_ooxml_parts) == expected_changed
    with ZipFile(BASE, "r") as before, ZipFile(output, "r") as after:
        assert before.namelist() == after.namelist()
        assert all(
            before.read(name) == after.read(name)
            for name in before.namelist()
            if name not in expected_changed
        )
        assert all(
            before.read(name) == after.read(name)
            for name in before.namelist()
            if name.endswith(".rels")
        )
        for protected_sheet in ("sheet1.xml", "sheet3.xml", "sheet5.xml"):
            part = f"xl/worksheets/{protected_sheet}"
            assert before.read(part) == after.read(part)
    assert _defined_names(output) == _defined_names(BASE)
    assert sha256_file(BASE) == EXPECTED_VALUATION_GOLDEN_SHA256


def test_preview_replay_is_raw_semantic_and_canonical_deterministic(
    tmp_path: Path, plan
) -> None:
    first = tmp_path / "preview_a.xlsx"
    second = tmp_path / "preview_b.xlsx"
    first_result = materialize_capital_return_debt_workbook_projection(
        plan=plan,
        base_workbook=BASE,
        output_workbook=first,
    )
    second_result = materialize_capital_return_debt_workbook_projection(
        plan=plan,
        base_workbook=BASE,
        output_workbook=second,
    )
    assert first.read_bytes() == second.read_bytes()
    assert first_result.output_workbook_sha256 == second_result.output_workbook_sha256
    assert first_result.canonical_ooxml_sha256 == second_result.canonical_ooxml_sha256
    assert canonical_ooxml_sha256(first) == canonical_ooxml_sha256(second)


def test_documentation_matches_executable_capital_return_ownership() -> None:
    manifest = (Path(__file__).parents[1] / "docs" / "standard_template_shell_manifest.md").read_text(
        encoding="utf-8"
    )
    assert "Capital Allocation at `A130:M143`" in manifest
    assert "Capital Return at `A145:M178`" in manifest
    assert "`Capital_Product_Lineage!A1:A30`" in manifest
    assert "`Valuation!A152:M168`" not in manifest
    assert "`Valuation!AD172:AO186`" not in manifest
