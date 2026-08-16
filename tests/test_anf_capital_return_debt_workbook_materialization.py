from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pytest

from pbi_xbrl.longitudinal_memory.capital_return_debt_workbook_materialization import (
    FormulaAwareCellMutation,
    WorkbookSheetStateMutation,
    WorksheetColumnMutation,
    WorksheetDimensionMutation,
    WorksheetRowMutation,
    WorksheetTableMutation,
    materialize_capital_return_debt_mutations,
)
from pbi_xbrl.longitudinal_memory.formula_aware_workbook_materialization import (
    FormulaAwareMaterializationError,
)


def _base_workbook(path: Path, *, nonempty_tail: bool = False) -> None:
    workbook = Workbook()
    valuation = workbook.active
    valuation.title = "Valuation"
    valuation["A1"] = 1
    valuation["B1"] = "legacy"
    valuation["E1"] = "retire"
    valuation["A5"] = "row"
    if nonempty_tail:
        valuation["A9"] = "must remain"
    product = workbook.create_sheet("HiddenProduct")
    product["A1"] = "Metric"
    product["B1"] = "Value"
    product["C1"] = "State"
    product["A1"].font = Font(bold=True, color="FFFFFF")
    for row in range(2, 7):
        product.cell(row, 1, f"metric-{row}")
        product.cell(row, 2, row)
        product.cell(row, 3, "available")
    table = Table(displayName="ProductTable", ref="A1:C6")
    table.autoFilter = AutoFilter(ref="A1:C6")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    product.add_table(table)
    workbook.save(path)
    workbook.close()


def test_bounded_extension_preserves_accepted_materializer_and_applies_structure(
    tmp_path: Path,
) -> None:
    base = tmp_path / "base.xlsx"
    output = tmp_path / "candidate.xlsx"
    _base_workbook(base)
    result = materialize_capital_return_debt_mutations(
        base_workbook=base,
        output_workbook=output,
        cell_mutations=(
            FormulaAwareCellMutation("Valuation", "B1", "SET_VALUE", "2", "number"),
            FormulaAwareCellMutation(
                "Valuation",
                "C1",
                "SET_VALUE",
                "header",
                "text",
                style_source_cell="A1",
                semantic_owner="bounded_presentation",
                style_source_sheet="HiddenProduct",
            ),
            FormulaAwareCellMutation("Valuation", "E1", "REMOVE_CELL"),
        ),
        row_mutations=(WorksheetRowMutation("Valuation", 5, hidden=True, height=30.0),),
        column_mutations=(WorksheetColumnMutation("Valuation", 1, 28.0),),
        dimension_mutations=(WorksheetDimensionMutation("Valuation", "A1:E5"),),
        sheet_state_mutations=(WorkbookSheetStateMutation("HiddenProduct", "hidden"),),
        table_mutations=(
            WorksheetTableMutation(
                "HiddenProduct",
                "A1:C5",
                ("Metric", "Value", "State"),
                show_row_stripes=False,
            ),
        ),
    )
    workbook = load_workbook(output, data_only=False)
    try:
        valuation = workbook["Valuation"]
        assert valuation["B1"].value == 2
        assert valuation["C1"].value == "header"
        assert valuation["C1"].style_id == workbook["HiddenProduct"]["A1"].style_id
        assert valuation["E1"].value is None
        assert valuation.row_dimensions[5].hidden is True
        assert valuation.row_dimensions[5].height == 30.0
        assert valuation.column_dimensions["A"].width == 28.0
        assert workbook["HiddenProduct"].sheet_state == "hidden"
        table = workbook["HiddenProduct"].tables["ProductTable"]
        assert table.ref == "A1:C5"
        assert table.tableStyleInfo.showRowStripes is False
    finally:
        workbook.close()
    with ZipFile(output, "r") as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml")
    assert b'r="E1"' not in sheet_xml
    assert result.cell_mutation_count == 3
    assert result.write_type_counts == {"number": 1, "remove": 1, "text": 1}
    assert result.column_mutation_count == 1
    assert result.dimension_mutation_count == 1
    assert result.sheet_state_mutation_count == 1
    assert result.table_mutation_count == 1


def test_dimension_tail_trim_is_fail_closed(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    output = tmp_path / "candidate.xlsx"
    _base_workbook(base, nonempty_tail=True)
    with pytest.raises(FormulaAwareMaterializationError, match="non-empty row"):
        materialize_capital_return_debt_mutations(
            base_workbook=base,
            output_workbook=output,
            cell_mutations=(),
            dimension_mutations=(
                WorksheetDimensionMutation("Valuation", "A1:E5", trim_empty_tail=True),
            ),
        )
