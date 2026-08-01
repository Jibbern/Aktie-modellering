from __future__ import annotations

import json
import hashlib
from html import unescape
import re
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pbi_xbrl.new_ticker_binding_planner import plan_standard_template_writes
from pbi_xbrl.standard_template_shell_identity import verify_shell_identity
from scripts.build_anf_shadow_normalized_package import build_anf_normalized_package


ROOT = Path(__file__).resolve().parents[1]


def _canonical_digest(value: object) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _data_root() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = parent / "StockModelData"
        if candidate.exists():
            return candidate
    return ROOT.parent / "StockModelData"


DATA_ROOT = _data_root()
ANF_WORKBOOK = DATA_ROOT / "outputs" / "Excel stock models" / "ANF_model.xlsx"
ANF_PACKAGE = DATA_ROOT / "outputs" / "stress_tests" / "ANF_new_ticker_engine" / "ANF_normalized_data_package.json"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
ANF_2019_RELEASE = DATA_ROOT / "tickers" / "ANF" / "earnings_release" / "8-K_2019-03-07_earnings_release.htm"


def _bindings() -> dict[str, dict]:
    return {
        binding["binding_id"]: binding
        for binding in json.loads(BINDING_MAP.read_text(encoding="utf-8"))["bindings"]
    }


def test_fy2018_cash_and_inventory_match_the_source_release_scale() -> None:
    assert ANF_2019_RELEASE.exists()
    source_html = ANF_2019_RELEASE.read_text(encoding="utf-8", errors="ignore")
    source_text = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", unescape(source_html)))
    assert re.search(
        r"Consolidated Balance Sheets \(in thousands\).*?Cash and equivalents \$ 723,135.*?Inventories 437,879",
        source_text,
    )

    wb = load_workbook(ANF_WORKBOOK, read_only=True, data_only=True)
    try:
        history = wb["History_Q"]
        headers = {str(history.cell(1, column).value or ""): column for column in range(1, history.max_column + 1)}
        row_number = next(
            row
            for row in range(2, history.max_row + 1)
            if str(history.cell(row, headers["fiscal_label"]).value or "") == "2018-Q4"
        )
        assert history.cell(row_number, headers["cash"]).value == 723_135_000_000
        assert history.cell(row_number, headers["inventory"]).value == 437_879_000_000
    finally:
        wb.close()

    package = build_anf_normalized_package(data_root=DATA_ROOT, workbook_path=ANF_WORKBOOK)
    annual = next(row for row in package["annual_financials"]["rows"] if row["period"] == "2018-FY")

    assert annual["cash"]["value"] == 723.135
    assert annual["inventory"]["value"] == 437.879
    assert "8-K_2019-03-07_earnings_release.htm" in annual["cash"]["source_ref"]
    assert "8-K_2019-03-07_earnings_release.htm" in annual["inventory"]["source_ref"]
    assert annual["cash"]["value"] + annual["inventory"]["value"] < annual["current_assets"]["value"]
    assert annual["current_assets"]["value"] < annual["total_assets"]["value"]


def test_anf_legacy_oracle_confirms_the_six_business_key_contracts_read_only() -> None:
    """ANF is a migration oracle, not input logic for generic onboarding."""

    assert ANF_WORKBOOK.exists()
    package = build_anf_normalized_package(data_root=DATA_ROOT, workbook_path=ANF_WORKBOOK)
    binding_payload = json.loads(BINDING_MAP.read_text(encoding="utf-8"))
    bindings = _bindings()
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    identity = verify_shell_identity(SHELL, manifest=manifest, binding_payload=binding_payload)
    plan = plan_standard_template_writes(
        package,
        binding_payload=binding_payload,
        manifest=manifest,
        shell_identity_report=identity,
    )
    assert plan.status == "PASS", [issue.to_dict() for issue in plan.issues]
    writes = {(write.target_sheet, write.target_cell): write for write in plan.planned_writes}

    wb = load_workbook(ANF_WORKBOOK, read_only=True, data_only=False)
    try:
        summary = wb["SUMMARY"]
        valuation = wb["Valuation"]
        segments = wb["BS_Segments"]
        drivers = wb["Operating_Drivers"]
        notes = wb["Quarter_Notes_UI"]
        promise = wb["Promise_Progress_UI"]
        investment_case = wb["ANF_Investment_Case"]

        # 1. Quarterly financials: ANF's original labels prove B26 is an as-of
        # value, while B28/B30 and Valuation's period/value rows are distinct.
        assert summary["A26"].value == "As of quarter"
        assert summary["A28"].value == "Revenue (latest quarter)"
        assert summary["A30"].value == "Net income (latest quarter)"
        assert valuation["A6"].value == "Quarter"
        assert valuation["A9"].value == "Revenue"
        assert valuation["A24"].value == "Adj EBITDA"
        assert valuation["A36"].value == "Net income attrib. to A&F"
        assert valuation["A43"].value == "CFO"
        for offset, row in enumerate(package["quarterly_financials"]["rows"], start=2):
            assert row["revenue"]["value"] == valuation.cell(9, offset).value
            assert row["adjusted_ebitda"]["value"] == valuation.cell(24, offset).value
            assert row["net_income"]["value"] == valuation.cell(36, offset).value
            assert row["operating_cash_flow"]["value"] == valuation.cell(43, offset).value
            assert "ANF_model.xlsx!History_Q" in row["revenue"]["source_ref"]
        assert package["quarterly_financials"]["rows"][8]["base_ebitda"]["value"] == valuation["J18"].value
        assert package["quarterly_financials"]["rows"][8]["adjusted_ebitda"]["value"] == valuation["J24"].value
        annuals = {row["period"]: row for row in package["annual_financials"]["rows"]}
        assert annuals["2025-FY"]["revenue"]["value"] == 5266.292
        assert annuals["2025-FY"]["adjusted_ebitda"]["value"] == 815.59
        assert annuals["2025-FY"]["net_income"]["value"] == 506.921
        assert annuals["2025-FY"]["free_cash_flow"]["value"] == 378.368
        assert annuals["2025-FY"]["operating_cash_flow"]["value"] == 619.142
        assert annuals["2018-FY"]["cash"]["value"] == 723.135
        assert annuals["2018-FY"]["inventory"]["value"] == 437.879
        assert ("BS_Segments", "B102") not in writes
        assert bindings["summary_as_of_quarter"]["source_field"] == "period"
        assert bindings["summary_latest_revenue"]["source_field"] == "revenue"
        assert bindings["summary_latest_net_income"]["source_field"] == "net_income"
        assert bindings["valuation_period_headers"]["source_field"] == "period"
        quarterly = {row["period"]: row for row in package["quarterly_financials"]["rows"]}
        latest = quarterly["2026-Q1"]
        prior_year = quarterly["2025-Q1"]
        trailing_periods = ("2025-Q2", "2025-Q3", "2025-Q4", "2026-Q1")

        def free_cash_flow(row: dict) -> float:
            return row["operating_cash_flow"]["value"] - row["capital_expenditures"]["value"]

        revenue_ttm = sum(quarterly[period]["revenue"]["value"] for period in trailing_periods)
        fcf_ttm = sum(free_cash_flow(quarterly[period]) for period in trailing_periods)
        latest_eps = latest["eps"]["value"]
        prior_eps = prior_year["eps"]["value"]
        operating_income_ttm = sum(
            quarterly[period]["operating_income"]["value"] for period in trailing_periods
        )
        interest_expense_ttm = sum(
            quarterly[period]["interest_expense"]["value"] for period in trailing_periods
        )

        assert writes[("SUMMARY", "B26")].value == "2026-Q1"
        assert writes[("SUMMARY", "B28")].value == pytest.approx(1113.821)
        assert writes[("SUMMARY", "B30")].value == pytest.approx(67.134)
        assert revenue_ttm == pytest.approx(5282.802)
        assert (latest["revenue"]["value"] - prior_year["revenue"]["value"]) / prior_year["revenue"]["value"] == pytest.approx(0.015045871225204177)
        assert (latest["net_income"]["value"] - prior_year["net_income"]["value"]) / prior_year["net_income"]["value"] == pytest.approx(-0.1651349906109708)
        assert latest_eps == pytest.approx(1.4697550189373207)
        assert latest_eps == pytest.approx(
            latest["net_income"]["value"] / latest["diluted_shares"]["value"]
        )
        assert (latest_eps - prior_eps) / abs(prior_eps) == pytest.approx(-0.07562577425325745)
        assert fcf_ttm == pytest.approx(416.047)
        assert (free_cash_flow(latest) - free_cash_flow(prior_year)) / abs(free_cash_flow(prior_year)) == pytest.approx(0.688024979913812)
        assert operating_income_ttm / abs(interest_expense_ttm) == pytest.approx(34.09362737793672)
        assert writes[("SUMMARY", "A3")].value == package["company_profile"]["business_description"]["value"]
        assert writes[("SUMMARY", "A5")].value == package["company_profile"]["strategic_context"]["value"]
        assert writes[("SUMMARY", "A3")].source_ref in package["company_profile"]["business_description"]["evidence_refs"]
        assert writes[("SUMMARY", "A5")].source_ref in package["company_profile"]["strategic_context"]["evidence_refs"]
        assert writes[("SUMMARY", "A9")].value == "Americas"
        assert writes[("SUMMARY", "B9")].value == 81.5

        # 2. Guidance: the legacy UI supplies the column semantics. Lineage
        # remains on planned writes; T9 is a non-anchor inside S9:Z9.
        assert (valuation["O8"].value, valuation["R8"].value, valuation["S8"].value) == (
            "Metric",
            "Applies to",
            "Guidance",
        )
        current_guidance = sorted(
            (item for item in package["normalized_guidance"]["items"] if item.get("display_role") == "current_primary"),
            key=lambda item: item["display_priority"],
        )
        assert len(current_guidance) == 7
        guidance = current_guidance[0]
        assert guidance["metric"]["source_ref"] == guidance["value"]["source_ref"]
        assert guidance["publication_date"] == "2026-03-04"
        assert guidance["source_date"] == "2026-01-31"
        assert guidance["stated_in_period"] == "2025-Q4"
        assert guidance["horizon"]["value"] in {"2026 year", "2026-Q1"}
        assert [(item["metric"]["value"], item["horizon"]["value"]) for item in current_guidance] == [
            ("Revenue", "2026 year"),
            ("Revenue", "2026-Q1"),
            ("Operating margin", "2026 year"),
            ("Operating margin", "2026-Q1"),
            ("Adj EPS", "2026 year"),
            ("Adj EPS", "2026-Q1"),
            ("Real estate activity", "2026 year"),
        ]
        guidance_sources = {
            column["source_field"]
            for column in bindings["valuation_guidance_current_primary_rows"]["target_columns"]
        }
        assert {
            "metric_display",
            "stated_period",
            "horizon",
            "value",
            "unit",
            "publication_date",
            "evidence_key",
            "display_state",
        } == guidance_sources
        assert "T" not in {
            column["target_column"]
            for column in bindings["valuation_guidance_current_primary_rows"]["target_columns"]
        }
        assert writes[("Valuation", "O9")].value == "Revenue"
        assert writes[("Valuation", "Q9")].value == "2025-Q4"
        assert writes[("Valuation", "R9")].value == "FY2026"
        assert writes[("Valuation", "X9")].value == "%"
        assert writes[("Valuation", "Y9")].value == "2026-03-04"
        assert writes[("Valuation", "Z9")].value == guidance["evidence_key"]
        assert writes[("Valuation", "AA9")].value == "current_primary / accepted"
        assert [
            (
                writes[("Valuation", f"O{row_idx}")].value,
                writes[("Valuation", f"R{row_idx}")].value,
                writes[("Valuation", f"S{row_idx}")].value,
                writes[("Valuation", f"X{row_idx}")].value,
                writes[("Valuation", f"Y{row_idx}")].value,
                writes[("Valuation", f"AA{row_idx}")].value,
            )
            for row_idx in range(9, 16)
        ] == [
            ("Revenue", "FY2026", "3%, 5%", "%", "2026-03-04", "current_primary / accepted"),
            ("Revenue", "2026-Q1", "1%, 3%", "%", "2026-03-04", "current_primary / accepted"),
            ("Operating margin", "FY2026", "12.0%, 12.5%", "%", "2026-03-04", "current_primary / accepted"),
            ("Operating margin", "2026-Q1", "around 7.0%", "%", "2026-03-04", "current_primary / accepted"),
            ("Adj EPS", "FY2026", "$10.20, $11.00", "$/share", "2026-03-04", "current_primary / accepted"),
            ("Adj EPS", "2026-Q1", "$1.20, $1.30", "$/share", "2026-03-04", "current_primary / accepted"),
            (
                "Real estate activity",
                "FY2026",
                "55 openings, 25 closures; 70 remodels/right-sizes",
                "stores",
                "2026-03-04",
                "current_primary / accepted",
            ),
        ]
        assert writes[("Promise_Progress_UI", "I61")].value == "2025-Q4"
        assert writes[("Promise_Progress_UI", "J61")].value == "2026-03-04"
        current_secondary = {
            (item["metric"]["value"], item["horizon"]["value"])
            for item in package["normalized_guidance"]["items"]
            if item.get("display_role") == "current_secondary"
        }
        assert {
            ("Capex", "2026 year"),
            ("Diluted shares", "2026 year"),
            ("Diluted shares", "2026-Q1"),
            ("Real estate activity", "2026-Q1"),
            ("Share repurchases", "2026 year"),
            ("Share repurchases", "2026-Q1"),
        } <= current_secondary
        assert [
            (
                writes[("Valuation", f"O{row_idx}")].value,
                writes[("Valuation", f"R{row_idx}")].value,
                writes[("Valuation", f"S{row_idx}")].value,
                writes[("Valuation", f"X{row_idx}")].value,
                writes[("Valuation", f"Y{row_idx}")].value,
                writes[("Valuation", f"AA{row_idx}")].value,
            )
            for row_idx in range(16, 22)
        ] == [
            ("Capex", "FY2026", "$200 million, $225 million", "$m", "2026-03-04", "current_secondary / accepted"),
            ("Diluted shares", "2026-Q1", "around 46 million", "shares_m", "2026-03-04", "current_secondary / accepted"),
            ("Diluted shares", "FY2026", "around 45 million", "shares_m", "2026-03-04", "current_secondary / accepted"),
            ("Real estate activity", "2026-Q1", "~30 net store openings", "stores", "2026-03-04", "current_secondary / accepted"),
            ("Share repurchases", "2026-Q1", "at least $100 million", "$m", "2026-03-04", "current_secondary / accepted"),
            ("Share repurchases", "FY2026", "around $450 million", "$m", "2026-03-04", "current_secondary / accepted"),
        ]
        assert [
            (
                writes[("Valuation", f"O{row_idx}")].value,
                writes[("Valuation", f"R{row_idx}")].value,
                writes[("Valuation", f"S{row_idx}")].value,
                writes[("Valuation", f"X{row_idx}")].value,
                writes[("Valuation", f"Y{row_idx}")].value,
                writes[("Valuation", f"AA{row_idx}")].value,
            )
            for row_idx in range(29, 36)
        ] == [
            ("Revenue", "FY2025", "at least 6%", "%", "2026-01-12", "history / accepted"),
            ("Operating margin", "FY2025", "around 13%", "%", "2026-01-12", "history / accepted"),
            ("Adj EPS", "FY2025", "$10.30, $10.40", "$/share", "2026-01-12", "history / accepted"),
            ("Capex", "FY2025", "~ $245 million", "$m", "2026-01-12", "history / accepted"),
            ("Diluted shares", "FY2025", "around 48 million", "shares_m", "2026-01-12", "history / accepted"),
            ("Real estate activity", "FY2025", "~40 net store openings", "stores", "2026-01-12", "history / accepted"),
            ("Share repurchases", "FY2025", "around $450 million", "$m", "2026-01-12", "history / accepted"),
        ]
        assert [
            (
                writes[("Valuation", f"O{row_idx}")].value,
                writes[("Valuation", f"X{row_idx}")].value,
                writes[("Valuation", f"Z{row_idx}")].value,
            )
            for row_idx in range(51, 59)
        ] == [
            ("Key debate", "manual_review_required", "investment_case.key_debate"),
            ("Why it can work", "accepted", "investment_case.why_it_can_work"),
            ("Upside factors", "manual_review_required", "investment_case.upside_factors"),
            ("Downside factors", "manual_review_required", "investment_case.downside_factors"),
            ("Watch next", "manual_review_required", "investment_case.watch_next"),
            ("Current stance", "manual_review_required", "investment_case.current_stance"),
            (
                "Sales-execution invalidator",
                "manual_review_required",
                "investment_case.invalidators.sales-execution-breaks",
            ),
            (
                "Margin-durability invalidator",
                "manual_review_required",
                "investment_case.invalidators.margin-durability-breaks",
            ),
        ]

        # 3. Segment rows: the generic contract uses neutral dimension slots.
        assert segments["A7"].value == "Quarter"
        assert segments["A50"].value == "ANF retail BS drivers"
        assert package["segments"]["items"]
        assert bindings["bs_segment_quarterly_rows"]["planning_state"] == "active"
        assert bindings["bs_segment_quarterly_rows"]["planning_mode"] == "pivot_rows"
        assert bindings["bs_segment_quarterly_rows"]["planner_target"] == "A61:M67"
        assert bindings["bs_segment_quarterly_rows"]["row_key"] == ["period", "dimension", "member", "metric"]
        hollister = next(item for item in package["segments"]["items"] if item["member"] == "Hollister" and item["metric"] == "revenue" and item["period"] == "2025-Q4")
        assert hollister["revenue"]["value"] == 863.3
        americas = next(item for item in package["segments"]["items"] if item["member"] == "Americas" and item["metric"] == "revenue" and item["period"] == "2025-FY")
        assert americas["annual_revenue"]["value"] == pytest.approx(4290.4, abs=0.01)
        assert [writes[("BS_Segments", f"{column}7")].value for column in "BCDEFGHIJKLM"] == [
            "2023-Q2",
            "2023-Q3",
            "2023-Q4",
            "2024-Q1",
            "2024-Q2",
            "2024-Q3",
            "2024-Q4",
            "2025-Q1",
            "2025-Q2",
            "2025-Q3",
            "2025-Q4",
            "2026-Q1",
        ]
        assert writes[("BS_Segments", "L66")].value == pytest.approx(863.3)
        assert writes[("BS_Segments", "L66")].row_key == "2025-Q4|brand|hollister|revenue"
        assert writes[("BS_Segments", "L61")].value == pytest.approx(segments["H61"].value)
        assert [writes[("BS_Segments", f"{column}70")].value for column in "BCDEFGHI"] == [
            "2018-FY",
            "2019-FY",
            "2020-FY",
            "2021-FY",
            "2022-FY",
            "2023-FY",
            "2024-FY",
            "2025-FY",
        ]
        assert writes[("BS_Segments", "G72")].value == pytest.approx(segments["B72"].value)
        assert writes[("BS_Segments", "I72")].value == pytest.approx(segments["D72"].value)

        # 4. Operating drivers preserve the generic three-column contract while
        # the ANF adapter supplies clean source-backed monitoring themes.
        assert drivers["A2"].value == "Operating Drivers"
        assert drivers["A5"].value == "Watch item"
        assert package["operating_drivers"]["items"]
        assert bindings["od_watchlist_rows"]["planning_state"] == "active"
        assert bindings["od_watchlist_rows"]["planner_target"] == "A6:N9"
        assert {column["target_column"] for column in bindings["od_watchlist_rows"]["target_columns"]} == {"A", "B", "H"}
        current_drivers = sorted((item for item in package["operating_drivers"]["items"] if item["display_role"] == "current_watchlist"), key=lambda item: item["display_priority"])
        assert len(current_drivers) == 4
        assert [item["topic"]["value"] for item in current_drivers] == [
            "Sales execution",
            "Margin durability",
            "Inventory quality",
            "Capital returns",
        ]
        for row_number, item in enumerate(current_drivers, start=6):
            assert writes[("Operating_Drivers", f"A{row_number}")].value == item["topic"]["value"]
            assert writes[("Operating_Drivers", f"B{row_number}")].value == item["current_read"]["value"]
            assert writes[("Operating_Drivers", f"H{row_number}")].value == item["why_it_matters"]["value"]
            assert item["current_read"]["source_ref"]
            assert item["current_read"]["source_ref"] == item["why_it_matters"]["source_ref"]

        # 5. Quarter notes retain legacy column intent without copying noisy
        # legacy UI prose or substituting Operating Drivers lineage.
        assert notes["C9"].value == "What happened"
        assert notes["H9"].value == "Model / valuation implication"
        assert notes["M9"].value == "Source / confidence"
        assert package["quarter_notes"]["items"]
        assert bindings["qn_quarter_note_rows"]["planning_state"] == "active"
        assert bindings["qn_quarter_note_rows"]["planner_target"] == "A10:M15"
        assert {column["target_column"] for column in bindings["qn_quarter_note_rows"]["target_columns"]} == {"A", "C", "F", "H", "M"}
        current_notes = sorted((item for item in package["quarter_notes"]["items"] if item["display_role"] == "current_note"), key=lambda item: item["display_priority"])
        assert len(current_notes) == 6
        assert [item["theme"]["value"] for item in current_notes] == [
            "Q4 results",
            "Brand mix",
            "Inventory",
            "2026 margin bridge",
            "Capital allocation",
            "Growth channels",
        ]
        for row_number, item in enumerate(current_notes, start=10):
            assert writes[("Quarter_Notes_UI", f"A{row_number}")].value == item["theme"]["value"]
            assert writes[("Quarter_Notes_UI", f"C{row_number}")].value == item["commentary"]["value"]
            assert writes[("Quarter_Notes_UI", f"F{row_number}")].value == item["why_it_matters"]["value"]
            assert writes[("Quarter_Notes_UI", f"H{row_number}")].value == item["model_implication"]["value"]
            assert writes[("Quarter_Notes_UI", f"M{row_number}")].value == item["source_display"]["value"]
            assert item["commentary"]["source_ref"] in item["evidence_refs"]
            visible_text = " ".join(str(item[field]["value"]) for field in ("commentary", "why_it_matters", "model_implication"))
            assert "Operating_Drivers" not in visible_text
            assert "binding" not in visible_text.casefold()
            assert "parser" not in visible_text.casefold()

        # 6. Investment case: the generic tokenized target is a scalar surface;
        # ANF's text remains read-only migration evidence, never template text.
        assert investment_case["A4"].value == "Investment Snapshot"
        assert investment_case["B5"].value
        assert package["investment_case"]["key_debate"]["evidence_classification"] == "analyst_interpretation_requiring_review"
        for cell, field in {
            "B5": "summary",
            "B6": "why_it_can_work",
            "B7": "key_debate",
            "B8": "upside_factors",
            "B9": "downside_factors",
            "B10": "watch_next",
            "B11": "current_stance",
        }.items():
            assert writes[("ANF_Investment_Case", cell)].value == package["investment_case"][field]["value"]
            assert package["investment_case"][field]["source_ref"]
        assert valuation["D198"].value is not None  # legacy display is an oracle, not source lineage
        assert package["valuation_inputs"]["net_debt"]["status"] == "missing_source"
        assert package["valuation_inputs"]["net_debt"]["value"] is None
        assert "D198 was not treated as evidence" in package["valuation_inputs"]["net_debt"]["reason"]
        assert package["valuation_inputs"]["base_ebitda_ttm"]["value"] == valuation["D199"].value
        assert package["valuation_inputs"]["adjusted_ebitda_ttm"]["value"] == valuation["D200"].value
        assert package["valuation_inputs"]["revenue_ttm"]["value"] == valuation["D203"].value
        assert package["debt_liquidity"]["total_debt"]["status"] == "missing_source"
        assert package["debt_liquidity"]["total_debt"]["value"] is None
        assert package["debt_liquidity"]["net_leverage"]["status"] == "missing_source"
        assert package["debt_liquidity"]["total_liquidity"]["value"] == pytest.approx(1209.086)
        assert package["debt_liquidity"]["as_of_date"]["value"] == "2026-01-31"
        assert package["debt_liquidity"]["liquidity_freshness"]["disposition"] == "stale_but_displayable_with_date"
        assert writes[("SUMMARY", "B45")].value == pytest.approx(1209.086)
        assert writes[("SUMMARY", "B45")].normalized_path == "debt_liquidity.summary_liquidity_display"
        assert writes[("SUMMARY", "D45")].value == "As of 2026-01-31 (stale)"
        assert writes[("SUMMARY", "B44")].value == pytest.approx(449.546)
        assert writes[("SUMMARY", "D44")].value == "As of 2026-01-31 (stale)"
        snapshot = {
            124: (594.08, "2026-05-02", "populated"),
            125: (449.546, "2026-01-31", "populated"),
            126: (1209.086, "2026-01-31", "populated"),
            127: (1292.477, "2026-05-02", "populated"),
        }
        for row_number, (value, period, status) in snapshot.items():
            assert writes[("Valuation", f"B{row_number}")].value == pytest.approx(value)
            assert writes[("Valuation", f"D{row_number}")].value == period
            assert writes[("Valuation", f"E{row_number}")].value == status
            assert writes[("Valuation", f"F{row_number}")].source_ref
        for row_number in range(128, 131):
            assert ("Valuation", f"B{row_number}") not in writes
            assert ("Valuation", f"D{row_number}") not in writes
            assert writes[("Valuation", f"E{row_number}")].value == "missing_source"
            assert writes[("Valuation", f"F{row_number}")].source_ref
        assert bindings["ic_investment_summary"]["planner_target"] == "B5"
        investment_summary = package["investment_case"]["summary"]
        assert investment_summary["source_ref"] in investment_summary["evidence_refs"]
        assert investment_summary["source_ref"].startswith("tickers/ANF/")
        assert investment_summary["evidence_classification"] == "evidence_backed_synthesis"

        # The old broad progression contract remains inactive. Typed current,
        # secondary, and historical rowsets own distinct visible blocks.
        assert [promise.cell(12, column).value for column in range(1, 10)] == [
            "Metric",
            "Initial guide",
            "Q1 update",
            "Q2 update",
            "Q3 update",
            "Q4 update",
            "Actual",
            "Status",
            "Notes/source",
        ]
        assert bindings["pp_annual_guidance_rows"]["planning_state"] == "inactive_legacy_contract"
        assert bindings["pp_progress_fy2025_rows"]["planning_state"] == "active"
        assert bindings["pp_progress_fy2024_rows"]["planning_state"] == "active"
        assert bindings["pp_current_secondary_guidance_rows"]["planning_state"] == "active"
        assert bindings["pp_guidance_timeline_rows"]["planning_state"] == "active"
        assert bindings["pp_guidance_timeline_rows"]["planner_target"] == "A61:K67"
        assert writes[("Promise_Progress_UI", "A13")].value == "FY2025 Revenue"
        assert writes[("Promise_Progress_UI", "A24")].value == "FY2024 Revenue"
        assert writes[("Promise_Progress_UI", "A39")].value == "Capex"
        assert writes[("Promise_Progress_UI", "C39")].value == "2026 year"
        assert bindings["summary_liquidity"]["planner_target"] == "B45"
        assert bindings["summary_net_leverage"]["planner_target"] == "B41"
        assert bindings["summary_net_leverage"]["normalized_field"] == "debt_liquidity.net_leverage"
        assert ("SUMMARY", "B41") not in writes
        leverage_gap = next(gap for gap in plan.mapping_gaps if gap.get("binding_id") == "summary_net_leverage")
        leverage_issue = next(issue for issue in plan.planner_issues if issue.binding_id == "summary_net_leverage")
        assert leverage_gap["source_ref"] == "ANF_model.xlsx!History_Q!row:50"
        assert leverage_issue.source_ref == leverage_gap["source_ref"]
    finally:
        wb.close()


def test_anf_planner_preserves_business_semantics_and_reconciles_final_qa_snapshot() -> None:
    assert ANF_WORKBOOK.exists()
    package = build_anf_normalized_package(data_root=DATA_ROOT, workbook_path=ANF_WORKBOOK)
    binding_payload = json.loads(BINDING_MAP.read_text(encoding="utf-8"))
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    identity = verify_shell_identity(SHELL, manifest=manifest, binding_payload=binding_payload)

    plan = plan_standard_template_writes(
        package,
        binding_payload=binding_payload,
        manifest=manifest,
        shell_identity_report=identity,
    )
    business_writes = [write for write in plan.planned_writes if write.target_sheet not in {"QA_Log", "Needs_Review", "QA_Checks"}]
    summary = plan.issue_ledger["summary"]
    qa_writes = Counter(write.target_sheet for write in plan.planned_writes if write.target_sheet in {"QA_Log", "Needs_Review", "QA_Checks"})
    qa_bindings = {
        binding["sheet"]: binding
        for binding in binding_payload["bindings"]
        if binding.get("source_policy") == "validation-output" and binding.get("planning_state") == "active"
    }

    assert plan.status == "PASS", [issue.to_dict() for issue in plan.issues]
    assert plan.qa_snapshot_status == "stable"
    assert not plan.has_blockers
    assert len({(write.target_sheet, write.target_cell) for write in business_writes}) == len(business_writes)
    hidden_value_binding_ids = {
        "hidden_value_base_rows",
        "hidden_value_audit_rows",
        "hidden_value_recompute_rows",
        "hidden_value_flags_rows",
        "hidden_value_valuation_rows",
    }
    product_pass_2a_binding_ids = {
        "summary_revolver_availability",
        "summary_revolver_availability_as_of",
        "valuation_debt_snapshot_cash_value",
        "valuation_debt_snapshot_cash_as_of",
        "valuation_debt_snapshot_cash_evidence",
        "valuation_debt_snapshot_cash_status",
        "valuation_debt_snapshot_revolver_value",
        "valuation_debt_snapshot_revolver_as_of",
        "valuation_debt_snapshot_revolver_evidence",
        "valuation_debt_snapshot_revolver_status",
        "valuation_debt_snapshot_liquidity_value",
        "valuation_debt_snapshot_liquidity_as_of",
        "valuation_debt_snapshot_liquidity_evidence",
        "valuation_debt_snapshot_liquidity_status",
        "valuation_debt_snapshot_leases_value",
        "valuation_debt_snapshot_leases_as_of",
        "valuation_debt_snapshot_leases_evidence",
        "valuation_debt_snapshot_leases_status",
        "valuation_debt_snapshot_core_debt_value",
        "valuation_debt_snapshot_core_debt_as_of",
        "valuation_debt_snapshot_core_debt_evidence",
        "valuation_debt_snapshot_core_debt_status",
        "valuation_debt_snapshot_net_debt_value",
        "valuation_debt_snapshot_net_debt_as_of",
        "valuation_debt_snapshot_net_debt_evidence",
        "valuation_debt_snapshot_net_debt_status",
        "valuation_debt_snapshot_net_leverage_value",
        "valuation_debt_snapshot_net_leverage_as_of",
        "valuation_debt_snapshot_net_leverage_evidence",
        "valuation_debt_snapshot_net_leverage_status",
    }
    product_pass_2b_binding_ids = {
        "valuation_guidance_current_primary_rows",
        "valuation_guidance_current_secondary_rows",
        "valuation_guidance_historical_rows",
        "valuation_thesis_debate_rows",
    }
    product_pass_3a2_binding_ids = {
        "debt_profile_resolved_rows",
        "revolver_history_resolved_rows",
        "revolver_history_companion_rows",
        "leverage_liquidity_resolved_rows",
        "leverage_liquidity_availability_rows",
        "leverage_liquidity_companion_rows",
        "debt_credit_notes_resolved_rows",
        "debt_maturity_ladder_resolved_rows",
    }
    focused_pass_b1_binding_ids = {"ic_product_projection_rows"}
    capital_return_binding_ids = {
        "valuation_capital_return_latest_quarter_header",
        "valuation_capital_return_ttm_header",
        "valuation_capital_return_annual_header",
        "valuation_capital_return_product_rows",
        "valuation_capital_return_support_rows",
    }
    accepted_business_writes = [
        write for write in business_writes if write.binding_id not in hidden_value_binding_ids
    ]
    business_sheets = {write.target_sheet for write in accepted_business_writes}
    assert business_sheets == {
        "SUMMARY",
        "Valuation",
        "BS_Segments",
        "Operating_Drivers",
        "ANF_Investment_Case",
        "ANF_Investment_Case_Data",
        "Quarter_Notes_UI",
        "Promise_Progress_UI",
        "History_Q",
        "Scenario_Driver_Assumptions",
        "Debt_Profile",
        "Revolver_History",
        "Leverage_Liquidity",
        "Debt_Credit_Notes",
    }
    by_binding = Counter(write.binding_id for write in business_writes)
    assert {
        binding_id: by_binding[binding_id]
        for binding_id in (
            "ic_bull_base_bear_rows",
            "ic_scenario_bridge_rows",
        )
    } == {
        "ic_bull_base_bear_rows": 220,
        "ic_scenario_bridge_rows": 102,
    }
    assert {
        binding_id: by_binding[binding_id]
        for binding_id in hidden_value_binding_ids
    } == {
        "hidden_value_base_rows": 700,
        "hidden_value_audit_rows": 107,
        "hidden_value_recompute_rows": 1_176,
        "hidden_value_flags_rows": 0,
        "hidden_value_valuation_rows": 0,
    }
    populated_snapshot_bindings = {
        binding_id: 1
        for binding_id in product_pass_2a_binding_ids
        if not binding_id.endswith(("core_debt_value", "core_debt_as_of", "net_debt_value", "net_debt_as_of", "net_leverage_value", "net_leverage_as_of"))
    }
    unavailable_snapshot_bindings = product_pass_2a_binding_ids - populated_snapshot_bindings.keys()
    assert {
        binding_id: by_binding[binding_id]
        for binding_id in product_pass_2a_binding_ids
    } == {
        **populated_snapshot_bindings,
        **{binding_id: 0 for binding_id in unavailable_snapshot_bindings},
    }
    payload = plan.to_dict()
    scenario_binding_ids = {
        "ic_bull_base_bear_rows",
        "ic_scenario_bridge_rows",
    }
    additive_binding_ids = (
        scenario_binding_ids
        | hidden_value_binding_ids
        | product_pass_2a_binding_ids
        | product_pass_2b_binding_ids
        | product_pass_3a2_binding_ids
        | focused_pass_b1_binding_ids
        | capital_return_binding_ids
    )
    assert len(payload["planned_writes"]) == 23_613
    # Two old raw guidance filters contributed 189 exclusions each, while four
    # deferred scalar-as-table bindings contributed one apiece. Resolved 2B
    # rowsets and B2 retirement replace 388 non-economic skips.
    assert payload["structured_skip_count"] == 2_006 == 2_394 - 388
    assert payload["overflow_count"] == 0
    product_pass_2a_writes = [
        write
        for write in payload["planned_writes"]
        if write.get("binding_id") in product_pass_2a_binding_ids
    ]
    assert len(product_pass_2a_writes) == 24
    assert _canonical_digest(product_pass_2a_writes) == "00246549f99ec0985bbe45a1ac3925e7dc4979527412448cb886fb09226a61af"
    product_pass_2b_writes = [
        write
        for write in payload["planned_writes"]
        if write.get("binding_id") in product_pass_2b_binding_ids
    ]
    assert Counter(write["binding_id"] for write in product_pass_2b_writes) == {
        "valuation_guidance_current_primary_rows": 56,
        "valuation_guidance_current_secondary_rows": 48,
        "valuation_guidance_historical_rows": 56,
        "valuation_thesis_debate_rows": 32,
    }
    assert len(product_pass_2b_writes) == 192
    assert _canonical_digest(product_pass_2b_writes) == "53f5e48096354e3b3f596b6a999d996a21794aa7b455ee07d589be2acdbda69a"
    product_pass_3a2_writes = [
        write
        for write in payload["planned_writes"]
        if write.get("binding_id") in product_pass_3a2_binding_ids
    ]
    assert len(product_pass_3a2_writes) == 389
    assert _canonical_digest(product_pass_3a2_writes) == (
        "774cc923de372f915599414a60dcd10d52832c746f6c6d73e3275caa7fcef57f"
    )
    focused_pass_b1_writes = [
        write
        for write in payload["planned_writes"]
        if write.get("binding_id") in focused_pass_b1_binding_ids
    ]
    assert len(focused_pass_b1_writes) == 761
    assert _canonical_digest(focused_pass_b1_writes) == (
        "24837ba4d3a5c297a053838ea5ed234a266b9682a5a5f299210864ee098d27b6"
    )
    capital_return_writes = [
        write
        for write in payload["planned_writes"]
        if write.get("binding_id") in capital_return_binding_ids
    ]
    assert len(capital_return_writes) == 240
    assert _canonical_digest(capital_return_writes) == (
        "9247ed248bcbeb2bcd88aeac8a0626e6a25cfae0205b7c37e2aa31da50837080"
    )
    assert len(payload["planned_writes"]) - 157 - 389 - 761 - 240 == 22_066
    non_scenario_writes = [
        write
        for write in payload["planned_writes"]
        if write.get("binding_id") not in additive_binding_ids
    ]
    non_scenario_binding_reports = [
        report
        for report in payload["bindings"]
        if report.get("binding_id") not in additive_binding_ids
    ]
    # The accepted 20,518-write nonadditive baseline loses 82 retired
    # BS_Segments writes and 690 obsolete QA cells. Product Pass 2A then adds
    # 138 QA cells for six explicit unavailable scalar/date dispositions.
    assert len(non_scenario_writes) == 19_702
    assert _canonical_digest(non_scenario_writes) == "e32df690c6de2c2dc05773ab8333ffcb598046e3d01a417a6969a3d22a18de9c"
    # B1 replaces five inactive legacy Investment Case bindings with one
    # active typed rowset binding, which is excluded above as a bounded addition.
    assert len(non_scenario_binding_reports) == 117
    assert _canonical_digest(non_scenario_binding_reports) == "4f2f3894d8368fa2de9df6495a3a68360b78d7cea540d51bd82498ec0e7fe341"

    segment_writes = [
        write
        for write in payload["planned_writes"]
        if write.get("binding_id") in {"bs_segment_quarterly_rows", "bs_segment_annual_rows"}
    ]
    assert len(segment_writes) == 100
    assert _canonical_digest(segment_writes) == "8d776a3c96daa72e2e68ba7926f9f82a152ea34f9d59f74bdf7f06cb73eae3d8"
    assert all("Total Company" not in str(write.get("row_key") or "") for write in segment_writes)
    assert any(write.get("row_key") == "total_company|total_company" for write in segment_writes)
    assert any(
        write.get("row_key") == "2023-Q4|total_company|total_company|revenue"
        for write in segment_writes
    )

    def outside_segment_product_surface(write: dict) -> bool:
        if write.get("target_sheet") in {"QA_Log", "Needs_Review", "QA_Checks"}:
            return False
        if write.get("target_sheet") != "BS_Segments":
            return True
        match = re.fullmatch(r"[A-Z]+(\d+)", str(write.get("target_cell") or ""))
        return match is None or not 61 <= int(match.group(1)) <= 104

    stable_business_writes = [
        write
        for write in payload["planned_writes"]
        if outside_segment_product_surface(write)
        and write.get("binding_id") not in product_pass_2a_binding_ids
        and write.get("binding_id") not in product_pass_2b_binding_ids
        and write.get("binding_id") not in product_pass_3a2_binding_ids
        and write.get("binding_id") not in focused_pass_b1_binding_ids
        and write.get("binding_id") not in capital_return_binding_ids
        and write.get("binding_id") not in hidden_value_binding_ids
    ]
    assert len(stable_business_writes) == 5_421
    assert _canonical_digest(stable_business_writes) == "cf7fbc32d43d92432fd365403d773edfe67bf059f38215effa2f291bed1f159f"
    all_stable_business_writes = [
        write
        for write in payload["planned_writes"]
        if outside_segment_product_surface(write)
        and write.get("binding_id") not in product_pass_2b_binding_ids
        and write.get("binding_id") not in product_pass_3a2_binding_ids
        and write.get("binding_id") not in focused_pass_b1_binding_ids
        and write.get("binding_id") not in capital_return_binding_ids
        and write.get("binding_id") not in hidden_value_binding_ids
    ]
    assert len(all_stable_business_writes) == 5_445
    assert _canonical_digest(all_stable_business_writes) == "6d44d17aa9b8ae97d16c3f898758746594d0b2b6c831d7a4599d9dca0abe55b0"
    assert _canonical_digest(payload["period_axes"]) == "88b9f00e07414ea100180a8f574e4ca3ab14088885107d888f75e1b143ec8818"
    assert _canonical_digest(payload["issue_ledger"]) == "6371c550feb51c5aea91f32bec18d393a65316788cf304351e500ad89799e8d6"
    assert _canonical_digest(payload["issue_ledger"]["issues"]) == "a5078093a353190340a27b0e3991234850ec4053d685b8a0a24a58ed749d7c83"
    assert _canonical_digest(payload["issue_ledger"]["occurrences"]) == "aa250df969959a18fff1cdc2a3f0bec03ba35cfd165444316b003f6a74186250"
    assert _canonical_digest(payload["mapping_gaps"]) == "7d1b871de8f591f38dbd05e1ace993a5fdada34e38510e58d1790c3aeb3bac96"
    assert _canonical_digest(payload["manual_review_flags"]) == "be52e1510ee9fd6a64388dd66f80f192b4c0c8a03879015e97df9babbab4cf9b"
    product_pass_2a_issues = [
        issue
        for issue in payload["issue_ledger"]["issues"]
        if issue.get("binding_id") in product_pass_2a_binding_ids
    ]
    assert len(product_pass_2a_issues) == 6
    assert {issue["occurrence_count"] for issue in product_pass_2a_issues} == {2}
    assert summary["canonical_unique_issue_count"] == 755
    assert summary["detailed_occurrence_count"] == 2_311
    assert by_binding["valuation_period_headers"] == 12
    assert by_binding["valuation_revenue_series"] == 12
    assert by_binding["valuation_net_income_series"] == 12
    assert by_binding["valuation_operating_cash_flow_series"] == 12
    assert "bs_annual_financial_period_headers" not in by_binding
    assert by_binding["bs_annual_revenue_series"] == 8
    assert summary["detailed_occurrence_count"] == len(plan.issue_ledger["occurrences"])
    assert summary["detailed_occurrence_count"] == len(plan.manual_review_flags) + len(plan.mapping_gaps) + len(plan.issues)
    assert summary["detailed_occurrence_count"] == sum(issue["occurrence_count"] for issue in plan.issue_ledger["issues"])
    assert summary["canonical_unique_issue_count"] == len(plan.issue_ledger["issues"])
    assert summary["actionable_issue_count"] == len(plan.issue_ledger["qa_presentation"]["needs_review_rows"])
    presentation_rows = {
        "QA_Log": plan.issue_ledger["qa_presentation"]["qa_log_rows"],
        "Needs_Review": plan.issue_ledger["qa_presentation"]["needs_review_rows"],
        "QA_Checks": plan.issue_ledger["qa_presentation"]["qa_check_rows"],
    }
    for sheet_name, rows in presentation_rows.items():
        mapped_fields = [column["source_field"] for column in qa_bindings[sheet_name]["target_columns"]]
        expected_cells = sum(
            1
            for row in rows
            for field in mapped_fields
            if row.get(field) not in (None, "")
        )
        assert qa_writes[sheet_name] == expected_cells
