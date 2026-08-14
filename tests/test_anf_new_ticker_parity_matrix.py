from __future__ import annotations

import copy
import hashlib
from collections import Counter
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from pbi_xbrl.json_schema_validation import load_json_strict, validate_json_schema
from pbi_xbrl.new_ticker_binding_planner import (
    DEFAULT_MANIFEST,
    reproduce_binding_plan,
)
from pbi_xbrl.path_config import (
    resolve_effective_data_root_from_ancestors,
    write_config_data_root,
)
from pbi_xbrl.standard_template_audit_freshness import _portable_file_sha256
from scripts.build_anf_new_ticker_parity_matrix import (
    GUIDANCE_MODULE_ID,
    GUIDANCE_PROJECTION_AUTHORITY,
    GUIDANCE_PROJECTION_ID,
    build_parity_matrix,
    _guidance_destination_lineage,
)


ROOT = Path(__file__).resolve().parents[1]
MATRIX = ROOT / "docs" / "anf_new_ticker_parity_matrix.json"
SCHEMA = ROOT / "docs" / "anf_new_ticker_parity_matrix.schema.json"
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"


DATA_ROOT_RESOLUTION = resolve_effective_data_root_from_ancestors(ROOT)
if DATA_ROOT_RESOLUTION.data_root is None:
    raise FileNotFoundError("No healthy registered StockModelData root is available for the parity fixture")
DATA_ROOT = DATA_ROOT_RESOLUTION.data_root
ANF_DIR = DATA_ROOT / "outputs" / "stress_tests" / "ANF_new_ticker_engine"
PACKAGE = ANF_DIR / "ANF_normalized_data_package.json"
LEGACY = DATA_ROOT / "outputs" / "Excel stock models" / "ANF_model.xlsx"
PRODUCT_V2_MANIFEST = ROOT / "tests" / "fixtures" / "promise_progress" / "anf_product_v2_golden_manifest.v1.json"
PRODUCT_V2_1_MANIFEST = ROOT / "tests" / "fixtures" / "promise_progress" / "anf_product_v2_1_golden_manifest.v1.json"


def _matrix() -> dict:
    return load_json_strict(MATRIX)


@lru_cache(maxsize=1)
def _fresh_plan() -> dict:
    return reproduce_binding_plan(
        load_json_strict(PACKAGE),
        binding_payload=load_json_strict(BINDING_MAP),
        manifest=load_json_strict(DEFAULT_MANIFEST),
        shell_path=SHELL,
    ).to_dict()


@lru_cache(maxsize=1)
def _fresh_matrix() -> dict:
    return build_parity_matrix(
        package=load_json_strict(PACKAGE),
        plan=_fresh_plan(),
        legacy_path=LEGACY,
        shell_path=SHELL,
        binding_path=BINDING_MAP,
    )


def _active_guidance_binding_ids(binding_document: dict) -> set[str]:
    return {
        str(binding["binding_id"])
        for binding in binding_document["bindings"]
        if binding.get("module_id") == GUIDANCE_MODULE_ID
        and binding.get("planning_state") == "active"
    }


def _target_cells(worksheet, target: str) -> tuple:
    min_col, min_row, max_col, max_row = range_boundaries(target)
    return tuple(
        cell
        for row in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        )
        for cell in row
    )


def test_binding_map_digest_is_checkout_eol_portable(tmp_path: Path) -> None:
    canonical = BINDING_MAP.read_bytes().replace(b"\r\n", b"\n").replace(b"\r", b"\n")
    variants = {
        "lf": canonical,
        "crlf": canonical.replace(b"\n", b"\r\n"),
        "mixed": canonical.replace(b"\n", b"\r\n", 1),
    }
    package = load_json_strict(PACKAGE)
    plan = _fresh_plan()
    matrices = []
    raw_hashes = set()

    for name, payload in variants.items():
        binding_path = tmp_path / name / "workbook_binding_map.json"
        binding_path.parent.mkdir()
        binding_path.write_bytes(payload)
        raw_hashes.add(hashlib.sha256(payload).hexdigest())
        matrices.append(
            build_parity_matrix(
                package=package,
                plan=plan,
                legacy_path=LEGACY,
                shell_path=SHELL,
                binding_path=binding_path.resolve(),
            )
        )

    assert matrices[0] == matrices[1] == matrices[2]
    binding_digest = matrices[0]["source_digests"]["binding_map_sha256"]
    assert binding_digest == _portable_file_sha256(BINDING_MAP)
    assert binding_digest == hashlib.sha256(canonical).hexdigest()
    assert len(raw_hashes) == 3
    assert binding_digest not in raw_hashes - {hashlib.sha256(canonical).hexdigest()}


def test_parity_matrix_is_schema_valid_and_has_unique_business_keys() -> None:
    matrix = _matrix()
    assert matrix["version"] == "1.4.0"
    assert validate_json_schema(matrix, load_json_strict(SCHEMA)) == []
    parity_ids = [row["parity_id"] for row in matrix["entries"]]
    assert len(parity_ids) == len(set(parity_ids))
    assert matrix["summary"]["entry_count"] == len(parity_ids)


def test_registered_data_root_resolution_is_secondary_worktree_independent(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    secondary_worktree = workspace / "Code.worktrees" / "secondary"
    data_root = workspace / "StockModelData"
    secondary_worktree.mkdir(parents=True)
    (data_root / "sec_cache").mkdir(parents=True)
    (data_root / "tickers").mkdir()
    write_config_data_root(workspace, data_root)

    result = resolve_effective_data_root_from_ancestors(secondary_worktree, env={})
    assert result.data_root == data_root.resolve()
    assert result.source == "config"


def test_parity_authorities_are_registered_and_not_transient_audit_outputs() -> None:
    for path in (PACKAGE, LEGACY):
        assert path.is_relative_to(DATA_ROOT)
        lowered_parts = {part.casefold() for part in path.parts}
        assert "audit" not in lowered_parts
        assert not any("candidate" in part or "rendered" in part for part in lowered_parts)

    source = Path(__file__).read_text(encoding="utf-8")
    assert "ROOT.parents" + "[2]" not in source
    assert "C:" + "\\\\Users\\\\" not in source


def test_product_golden_fixture_versions_remain_isolated_from_parity_schema() -> None:
    product_v2 = load_json_strict(PRODUCT_V2_MANIFEST)
    product_v2_1 = load_json_strict(PRODUCT_V2_1_MANIFEST)

    assert product_v2["product_version"] == "2.0.0-candidate"
    assert product_v2_1["product_version"] == "2.1.0"
    product_v2_1_sha = next(
        artifact["sha256"]
        for artifact in product_v2_1["fixture_artifacts"]
        if artifact["relative_path"].endswith("anf_product.v2_1.json")
    )
    assert product_v2["product_sha256"] != product_v2_1_sha
    assert _matrix()["version"] == "1.4.0"


def test_all_available_required_items_are_reproduced() -> None:
    matrix = _matrix()
    reproduced_statuses = {"reproduced_correctly", "reproduced_with_improved_wording"}
    missing = [
        row
        for row in matrix["entries"]
        if row["parity_requirement"] == "must_reproduce"
        and row["current_status"] not in reproduced_statuses
    ]
    assert missing == []
    assert matrix["summary"]["required_missing_count"] == 0
    assert {
        row["inventory_origin"]
        for row in matrix["entries"]
        if row["parity_requirement"] == "must_reproduce"
    } <= {
        "legacy_workbook_business_key",
        "legacy_visible_display_contract",
        "source_evidence_business_key",
    }
    assert matrix["summary"]["independent_source_fact_reproduced_count"] == matrix["summary"]["independent_source_fact_count"]


def test_guidance_destination_lineage_exactly_matches_the_fresh_plan() -> None:
    plan = _fresh_plan()
    binding_document = load_json_strict(BINDING_MAP)
    active_binding_ids = _active_guidance_binding_ids(binding_document)
    planned_writes = [
        write
        for write in plan["planned_writes"]
        if write["binding_id"] in active_binding_ids
    ]
    planned_destinations = {
        f"{write['target_sheet']}!{write['target_cell']}"
        for write in planned_writes
    }

    lineage = _matrix()["summary"]["guidance_destination_lineage"]
    lineage_destinations = [
        destination["destination"]
        for binding in lineage["bindings"]
        for destination in binding["destinations"]
    ]
    lineage_by_id = {
        binding["binding_id"]: binding
        for binding in lineage["bindings"]
    }

    assert lineage["module_id"] == GUIDANCE_MODULE_ID
    assert lineage["module_profile_id"] == "full_union"
    assert lineage["active_binding_count"] == len(active_binding_ids) == 7
    assert lineage["destination_count"] == len(planned_destinations) == 320
    assert len(lineage_destinations) == len(set(lineage_destinations))
    assert set(lineage_destinations) == planned_destinations
    assert set(lineage_by_id) == active_binding_ids
    assert {"Valuation", "Promise_Progress_UI"} == {
        destination.split("!", 1)[0]
        for destination in lineage_destinations
    }

    for binding_id in (
        "valuation_guidance_current_primary_rows",
        "valuation_guidance_current_secondary_rows",
        "valuation_guidance_historical_rows",
    ):
        binding = lineage_by_id[binding_id]
        assert binding["source_selector_type"] == "derived_resolved_rowset"
        assert binding["normalized_collection_root"] == "normalized_guidance.items"
        assert binding["resolved_rowset_producer"] == GUIDANCE_PROJECTION_ID
        assert binding["resolver_projection_authority"] == GUIDANCE_PROJECTION_AUTHORITY
        assert binding["formula_or_value_ownership"] == "value_binding"

    for binding_id in (
        "pp_progress_fy2025_rows",
        "pp_progress_fy2024_rows",
        "pp_current_secondary_guidance_rows",
        "pp_guidance_timeline_rows",
    ):
        binding = lineage_by_id[binding_id]
        assert binding["source_selector_type"] == "direct_package_collection"
        assert not binding["resolved_rowset_producer"]
        assert not binding["resolver_projection_authority"]

    inactive_ids = {
        str(binding["binding_id"])
        for binding in binding_document["bindings"]
        if binding.get("module_id") == GUIDANCE_MODULE_ID
        and binding.get("planning_state") != "active"
    }
    assert inactive_ids
    assert inactive_ids.isdisjoint(lineage_by_id)
    assert _matrix() == _fresh_matrix()


def test_guidance_lineage_is_independent_of_plan_order_and_ticker_identity() -> None:
    plan = copy.deepcopy(_fresh_plan())
    binding_document = load_json_strict(BINDING_MAP)
    expected = _guidance_destination_lineage(plan, binding_document)

    plan["planned_writes"].reverse()
    plan["ticker"] = "XYZ"

    assert _guidance_destination_lineage(plan, binding_document) == expected


def test_derived_valuation_guidance_restores_the_exact_lost_destinations() -> None:
    rows = {row["parity_id"]: row for row in _matrix()["entries"]}
    expected = {
        "legacy-guidance:79:revenue:2026-Q1": "Valuation!S10",
        "legacy-guidance:80:revenue:FY2026": "Valuation!S9",
        "legacy-guidance:81:operating_margin:2026-Q1": "Valuation!S12",
        "legacy-guidance:82:operating_margin:FY2026": "Valuation!S11",
        "legacy-guidance:83:adjusted_eps:2026-Q1": "Valuation!S14",
        "legacy-guidance:84:adjusted_eps:FY2026": "Valuation!S13",
        "legacy-guidance:91:real_estate_activity:FY2026": "Valuation!S15",
        "promise-progress:adjusted_eps:FY2026": "Valuation!S13",
        "promise-progress:operating_margin:FY2026": "Valuation!S11",
        "promise-progress:real_estate_activity:FY2026": "Valuation!S15",
        "promise-progress:revenue:FY2026": "Valuation!S9",
    }

    for parity_id, destination in expected.items():
        row = rows[parity_id]
        assert destination in row["expected_new_workbook_destination"]
        assert "valuation_guidance_current_primary_rows" in row["binding_ids"]

    primary_value_destinations = {
        destination["destination"]
        for binding in _matrix()["summary"]["guidance_destination_lineage"]["bindings"]
        if binding["binding_id"] == "valuation_guidance_current_primary_rows"
        for destination in binding["destinations"]
        if destination["target_role"].endswith(".value")
    }
    assert primary_value_destinations == {
        "Valuation!S9",
        "Valuation!S10",
        "Valuation!S11",
        "Valuation!S12",
        "Valuation!S13",
        "Valuation!S14",
        "Valuation!S15",
    }


def test_promise_progress_legacy_aliases_and_visible_row_routes_are_reproduced() -> None:
    rows = {
        row["parity_id"]: row
        for row in _matrix()["entries"]
        if row["domain"] == "promise_progress"
    }
    expected_fy2025 = {
        "promise-progress:revenue:FY2025",
        "promise-progress:operating_margin:FY2025",
        "promise-progress:adjusted_eps:FY2025",
        "promise-progress:capital_expenditures:FY2025",
        "promise-progress:diluted_shares:FY2025",
        "promise-progress:real_estate_activity:FY2025",
        "promise-progress:share_repurchases:FY2025",
        "promise-progress:tariffs:FY2025",
    }
    assert expected_fy2025 <= rows.keys()
    for parity_id in expected_fy2025:
        row = rows[parity_id]
        assert row["current_status"] == "reproduced_correctly", parity_id
        assert row["disposition"] == "visible", parity_id
        assert row["expected_new_workbook_destination"], parity_id
        assert all(
            destination.startswith("Promise_Progress_UI!")
            for destination in row["expected_new_workbook_destination"]
        ), parity_id

    old_revenue = rows["promise-progress:revenue:FY2020"]
    assert old_revenue["current_status"] == "explicitly_rejected_with_evidence"
    assert old_revenue["disposition"] == "rejected_with_evidence"
    assert old_revenue["dimensions"]["legacy_occurrence_count"] == 1
    assert old_revenue["dimensions"]["source_refs"]
    assert "reopened-store sales productivity" in old_revenue["rejection_reason"]
    assert old_revenue["expected_new_workbook_destination"] == []


def test_promise_progress_parity_reports_explicit_key_and_occurrence_dispositions() -> None:
    matrix = _matrix()
    rows = {
        row["parity_id"]: row
        for row in matrix["entries"]
        if row["domain"] == "promise_progress"
    }

    assert matrix["summary"]["promise_progress_key_disposition_counts"] == {
        "audit_only": 6,
        "duplicate_superseded": 1,
        "rejected_with_evidence": 5,
        "visible_reproduced": 17,
    }
    assert matrix["summary"]["promise_progress_occurrence_disposition_counts"] == {
        "audit_only_historical_evidence": 9,
        "duplicate_or_superseded_evidence": 5,
        "rejected_with_evidence": 12,
    }
    assert rows["promise-progress:capital_expenditures:FY2020"]["current_status"] == "audit_only_evidence_preserved"
    assert rows["promise-progress:operating_margin:FY2023"]["current_status"] == "duplicate_or_superseded_evidence_preserved"
    assert rows["promise-progress:tariffs:FY2019"]["current_status"] == "explicitly_rejected_with_evidence"
    assert not any(row["disposition"] == "missing" for row in rows.values())


def test_parity_inventory_is_legacy_first_and_keeps_fy2018_fy2019_and_older_history() -> None:
    matrix = _matrix()
    assert matrix["inventory_method"].startswith("legacy workbook business keys are inventoried first")
    annual_rows = [
        row
        for row in matrix["entries"]
        if row["parity_id"].startswith("legacy-annual:") and row["inventory_class"] == "source_fact"
    ]
    assert {"2018-FY", "2019-FY"} <= {row["period"] for row in annual_rows}
    assert {"2015-FY", "2016-FY", "2017-FY"} <= {row["period"] for row in annual_rows}
    assert all(row["inventory_origin"] == "legacy_workbook_business_key" for row in annual_rows)

    for period in ("2018-FY", "2019-FY"):
        revenue = next(
            row
            for row in annual_rows
            if row["period"] == period and row["metric_business_meaning"] == "revenue"
        )
        assert revenue["legacy_sheet_range"].startswith("ANF_model.xlsx!History_Q!")
        assert revenue["comparison_result"] == "value_match"

    older = [row for row in annual_rows if row["period"] in {"2015-FY", "2016-FY", "2017-FY"}]
    assert older
    assert all(row["disposition"] in {"audit_only", "formula_owned", "explicitly_excluded"} for row in older)
    assert all(not row["expected_new_workbook_destination"] for row in older)


def test_removing_an_annual_package_row_cannot_remove_the_legacy_parity_item() -> None:
    package = copy.deepcopy(load_json_strict(PACKAGE))
    package["annual_financials"]["rows"] = [
        row for row in package["annual_financials"]["rows"] if row["period"] != "2018-FY"
    ]
    matrix = build_parity_matrix(
        package=package,
        plan=_fresh_plan(),
        legacy_path=LEGACY,
        shell_path=SHELL,
        binding_path=BINDING_MAP,
    )
    row = next(
        row
        for row in matrix["entries"]
        if row["parity_id"] == "legacy-annual:2018-FY:revenue"
    )
    assert row["normalized_package_path"] == "annual_financials.rows[missing:2018-FY].revenue"
    assert row["comparison_result"] == "missing_normalized_fact"
    assert row["current_status"] == "missing_or_explicitly_unavailable"


def test_quarterly_and_annual_core_financial_minimums_are_locked() -> None:
    matrix = _matrix()
    reproduced = Counter(
        (row["domain"], row["metric_business_meaning"])
        for row in matrix["entries"]
        if row["current_status"] == "reproduced_correctly"
    )
    for metric in (
        "revenue",
        "gross_profit",
        "operating_income",
        "base_ebitda",
        "adjusted_ebitda",
        "net_income",
    ):
        assert reproduced[("quarterly_financials", metric)] >= 12, metric
    for metric in ("operating_cash_flow", "capital_expenditures"):
        assert reproduced[("cash_flow", metric)] >= 12, metric
    assert reproduced[("per_share", "diluted_shares")] >= 12
    assert reproduced[("per_share", "eps")] >= 9
    assert reproduced[("per_share", "adjusted_eps")] >= 11
    for metric in ("revenue", "gross_profit", "operating_income", "net_income"):
        assert reproduced[("annual_financials", metric)] >= 8, metric
    for metric in ("base_ebitda", "operating_cash_flow", "capital_expenditures"):
        assert reproduced[("annual_financials", metric)] >= 6, metric
    assert reproduced[("annual_financials", "adjusted_ebitda")] >= 2
    assert reproduced[("annual_financials", "diluted_shares")] == 0
    assert reproduced[("annual_financials", "eps")] == 0


def test_source_backed_required_items_have_lineage_and_exact_destinations() -> None:
    for row in _matrix()["entries"]:
        if row["parity_requirement"] != "must_reproduce" or row["source_backed_vs_derived"] != "source_backed":
            continue
        assert row["source_ref"], row["parity_id"]
        assert row["normalized_package_path"], row["parity_id"]
        if not row["expected_new_workbook_destination"]:
            assert row["disposition"] in {
                "audit_only",
                "formula_owned",
                "explicitly_excluded",
                "history",
                "superseded",
            }, row["parity_id"]


def test_valuation_input_parity_covers_actual_optional_and_user_input_contracts() -> None:
    rows = {
        row["normalized_package_path"]: row
        for row in _matrix()["entries"]
        if row["domain"] == "valuation_inputs"
    }
    retired_duplicate_inputs = {
        "valuation_inputs.as_of_date",
        "valuation_inputs.diluted_shares",
        "valuation_inputs.base_ebitda_ttm",
        "valuation_inputs.adjusted_ebitda_ttm",
        "valuation_inputs.revenue_ttm",
        "valuation_inputs.operating_cash_flow_ttm",
        "valuation_inputs.free_cash_flow_ttm",
        "valuation_inputs.capex_ttm",
        "valuation_inputs.eps_ttm",
    }
    for path in retired_duplicate_inputs:
        assert rows[path]["parity_requirement"] == "intentionally_rejected"
        assert rows[path]["current_status"] == "explicitly_rejected_with_evidence"
        assert rows[path]["disposition"] == "duplicate_display_binding"
        assert rows[path]["expected_new_workbook_destination"] == []
        assert "B2 retired the duplicate Valuation display binding" in rows[path]["rejection_reason"]

    matrix_rows = {row["parity_id"]: row for row in _matrix()["entries"]}
    assert matrix_rows["formula:operating_cash_flow_ttm:2026-Q1"]["expected_new_workbook_destination"] == [
        "Valuation!M271"
    ]
    assert matrix_rows["formula:free_cash_flow_ttm:2026-Q1"]["expected_new_workbook_destination"] == [
        "Valuation!M49"
    ]

    for path in (
        "valuation_inputs.shares_outstanding",
        "valuation_inputs.net_debt",
        "valuation_inputs.adjusted_eps_ttm",
        "valuation_inputs.book_value_per_share",
        "valuation_inputs.tangible_book_value_per_share",
        "valuation_inputs.interest_paid_ttm",
    ):
        assert rows[path]["parity_requirement"] == "unavailable_missing_evidence"
        assert rows[path]["current_status"] == "missing_or_explicitly_unavailable"

    for path in (
        "valuation_inputs.price",
        "valuation_inputs.adjusted_fcf_ttm",
        "valuation_inputs.target_ev_adjusted_ebitda",
        "valuation_inputs.target_ev_ebitda",
        "valuation_inputs.target_ev_yield",
        "valuation_inputs.maintenance_capex_ratio",
        "valuation_inputs.recurring_cash_costs",
        "valuation_inputs.working_capital_normalization",
        "valuation_inputs.per_share_denominator",
    ):
        assert rows[path]["parity_requirement"] == "intentionally_rejected"
        assert rows[path]["current_status"] == "missing_or_explicitly_unavailable"
        assert rows[path]["rejection_reason"]


def test_formula_improvements_exist_in_protected_cells() -> None:
    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        for row in _matrix()["entries"]:
            if row["inventory_class"] != "formula_improvement":
                continue
            assert len(row["expected_new_workbook_destination"]) == 1
            sheet, coordinate = row["expected_new_workbook_destination"][0].split("!", 1)
            cells = _target_cells(wb[sheet], coordinate)
            assert cells
            assert all(isinstance(cell.value, str) and cell.value.startswith("=") for cell in cells), row["parity_id"]
            assert all(cell.protection.locked is True for cell in cells)
            assert row["formula_contract_status"] == "present_protected"
            assert row["economic_calculability"] in {
                "economically_calculable",
                "blank_due_to_missing_evidence",
            }
            assert row["calculation_reason"]
    finally:
        wb.close()

    formula_rows = [row for row in _matrix()["entries"] if row["inventory_class"] == "formula_improvement"]
    assert any(row["economic_calculability"] == "economically_calculable" for row in formula_rows)
    assert any(row["economic_calculability"] == "blank_due_to_missing_evidence" for row in formula_rows)
    assert all(
        row["current_status"]
        == (
            "reproduced_correctly"
            if row["economic_calculability"] == "economically_calculable"
            else "contract_present_blank_by_missing_evidence"
        )
        for row in formula_rows
    )


def test_legacy_cogs_tax_da_and_operating_margin_are_explicitly_classified() -> None:
    matrix = _matrix()
    for metric in ("cost_of_goods_sold", "income_taxes_paid", "depreciation_amortization"):
        rows = [
            row
            for row in matrix["entries"]
            if row["metric_business_meaning"] == metric and row["inventory_class"] == "source_fact"
        ]
        assert rows, metric
        assert all(row["inventory_origin"] == "legacy_workbook_business_key" for row in rows)
        assert all(row["disposition"] == "audit_only" for row in rows)

    margin_rows = [
        row
        for row in matrix["entries"]
        if row["parity_id"].startswith("legacy-quarter:")
        and row["metric_business_meaning"] == "operating_margin"
    ]
    assert margin_rows
    assert all(row["inventory_class"] == "source_fact" for row in margin_rows)
    assert all(row["source_backed_vs_derived"] == "derived" for row in margin_rows)
    assert all(row["disposition"] == "formula_owned" for row in margin_rows)
    assert all(row["formula_contract_status"] == "not_applicable" for row in margin_rows)

    formula_rows = [
        row
        for row in matrix["entries"]
        if row["parity_id"].startswith("formula:operating_margin:")
    ]
    assert formula_rows
    assert all(row["formula_contract_status"] == "present_protected" for row in formula_rows)


def test_missing_fail_zero_placeholders_are_unavailable_and_make_formulas_blank() -> None:
    matrix = _matrix()
    by_id = {row["parity_id"]: row for row in matrix["entries"]}
    unsupported_quarterly = {
        "total_debt": {"2024-Q2", "2024-Q3", "2024-Q4"},
        "debt_core": {"2024-Q2", "2024-Q3", "2024-Q4", "2025-Q2", "2025-Q3", "2025-Q4"},
        "interest_paid": {"2024-Q3", "2024-Q4", "2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"},
    }
    for metric, periods in unsupported_quarterly.items():
        for period in periods:
            row = by_id[f"legacy-quarter:{period}:{metric}"]
            assert row["parity_requirement"] == "unavailable_missing_evidence"
            assert row["inventory_class"] == "unsupported_legacy_content"
            assert row["inventory_origin"] == "legacy_report_quality_check"
            assert row["current_status"] == "missing_or_explicitly_unavailable"
            assert row["comparison_result"] == "unsupported_zero_placeholder_left_blank"
            assert row["expected_new_workbook_destination"] == []
            assert "Source=Missing or QA=FAIL" in row["rejection_reason"]

    for period, metric in (
        ("2024-FY", "total_debt"),
        ("2024-FY", "debt_core"),
        ("2024-FY", "interest_paid"),
        ("2025-FY", "debt_core"),
        ("2025-FY", "interest_paid"),
    ):
        row = by_id[f"legacy-annual:{period}:{metric}"]
        assert row["parity_requirement"] == "unavailable_missing_evidence"
        assert row["inventory_class"] == "unsupported_legacy_content"
        assert row["comparison_result"] == "unsupported_zero_placeholder_left_blank"
        assert row["expected_new_workbook_destination"] == []

    for period in ("2024-Q2", "2024-Q3", "2024-Q4", "2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"):
        if f"formula:net_debt:{period}" in by_id:
            assert by_id[f"formula:net_debt:{period}"]["economic_calculability"] == "blank_due_to_missing_evidence"
    for period in ("2024-Q3", "2024-Q4", "2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"):
        assert by_id[f"formula:cash_interest_coverage:{period}"]["economic_calculability"] == "blank_due_to_missing_evidence"
    assert not any(parity_id.startswith("formula:annual_net_debt:") for parity_id in by_id)


def test_segment_plan_preserves_dimension_identity() -> None:
    plan = _fresh_plan()
    labels = {
        str(write["value"])
        for write in plan["planned_writes"]
        if write["binding_id"] in {"bs_segment_quarterly_rows", "bs_segment_annual_rows"}
        and str(write["target_cell"]).startswith("A")
    }
    assert {"Geography: Americas", "Geography: EMEA", "Geography: APAC"} <= labels
    assert {"Brand: Hollister", "Brand: Abercrombie", "Total Company"} <= labels


def test_segment_parity_is_exactly_inventoried_from_legacy_visible_cells() -> None:
    wb = load_workbook(LEGACY, read_only=True, data_only=True)
    try:
        ws = wb["BS_Segments"]
        expected = set()
        for period_type, header_row, member_rows, columns in (
            ("quarterly", 7, (61, 62, 63, 65, 66, 67), range(2, 14)),
            ("annual", 70, (72, 73, 74), range(2, 10)),
        ):
            for row_number in member_rows:
                member = str(ws.cell(row_number, 1).value or "")
                dimension = "geography" if member in {"Americas", "EMEA", "APAC"} else "brand" if member in {"Hollister", "Abercrombie"} else "total_company"
                for column in columns:
                    period = ws.cell(header_row, column).value
                    value = ws.cell(row_number, column).value
                    if period in (None, "") or not isinstance(value, (int, float)) or isinstance(value, bool):
                        continue
                    normalized_period = f"{int(period)}-FY" if period_type == "annual" else str(period)
                    expected.add((normalized_period, dimension, member, float(value)))
    finally:
        wb.close()

    actual = {
        (
            row["period"],
            row["dimensions"]["dimension"],
            row["dimensions"]["member"],
            float(row["legacy_value"]),
        )
        for row in _matrix()["entries"]
        if row["domain"] == "segments" and row["inventory_origin"] == "legacy_workbook_business_key"
    }
    assert len(expected) == 52
    assert actual == expected


def test_annual_eps_and_share_proxies_are_explicitly_rejected() -> None:
    rows = [
        row
        for row in _matrix()["entries"]
        if row["parity_id"].startswith("legacy-annual:")
        and row["metric_business_meaning"] in {"diluted_shares", "eps"}
    ]
    annual_periods = {
        row["period"]
        for row in _matrix()["entries"]
        if row["parity_id"].startswith("legacy-annual:")
    }
    assert len(rows) == 2 * len(annual_periods)
    assert {row["metric_business_meaning"] for row in rows} == {"diluted_shares", "eps"}
    assert all(row["parity_requirement"] == "unavailable_missing_evidence" for row in rows)
    assert all("Q4" in row["rejection_reason"] for row in rows)


def test_generic_formula_and_planner_modules_contain_no_anf_business_logic() -> None:
    generic_paths = (
        ROOT / "pbi_xbrl" / "standard_template_formula_contract.py",
        ROOT / "pbi_xbrl" / "new_ticker_binding_planner.py",
        ROOT / "pbi_xbrl" / "new_ticker_value_filler.py",
    )
    for path in generic_paths:
        source = path.read_text(encoding="utf-8")
        assert "ANF_model.xlsx" not in source
        assert "Abercrombie" not in source
        assert "Hollister" not in source
