from __future__ import annotations

import copy
import json
import re
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pbi_xbrl.new_ticker_binding_planner import (
    BindingPlan,
    BindingPlanReproductionError,
    reproduce_binding_plan_snapshot,
)
from pbi_xbrl.normalized_company_data_validation import (
    build_normalized_text_quality_audit,
    validate_normalized_company_data,
)
from pbi_xbrl.standard_template_shell_identity import verify_shell_identity
from scripts.build_anf_shadow_normalized_package import (
    _annual_component_field,
    _annual_incomplete_candidate_reviews,
    _build_annual_financial_rows,
    _build_debt_liquidity,
    _build_quarterly_financial_rows,
    _build_valuation_inputs,
    _payload_sha256,
    _ttm_component_field,
    build_binding_coverage_audit,
    build_anf_normalized_package,
    build_anf_shadow_outputs,
)


ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = next(
    ancestor / "StockModelData"
    for ancestor in [ROOT, *ROOT.parents]
    if (ancestor / "StockModelData").exists()
)
ANF_WORKBOOK = DATA_ROOT / "outputs" / "Excel stock models" / "ANF_model.xlsx"
ANF_STRESS_DIR = DATA_ROOT / "outputs" / "stress_tests" / "ANF_new_ticker_engine"
ANF_PACKAGE = ANF_STRESS_DIR / "ANF_normalized_data_package.json"
ANF_PLAN = ANF_STRESS_DIR / "ANF_binding_plan.json"


def test_anf_shadow_package_reports_are_built_from_read_only_legacy_artifacts(tmp_path: Path) -> None:
    output_dir = tmp_path / "ANF_new_ticker_engine"

    paths = build_anf_shadow_outputs(
        data_root=DATA_ROOT,
        workbook_path=ANF_WORKBOOK,
        output_dir=output_dir,
        docs_dir=tmp_path / "docs",
    )

    package_path = paths["package"]
    mapping_path = paths["mapping_gaps"]
    validation_path = paths["validation"]
    source_audit_path = paths["source_audit_json"]
    coverage_path = paths["binding_coverage_json"]
    text_quality_path = paths["text_quality_json"]

    for path in (package_path, mapping_path, validation_path, source_audit_path, coverage_path, text_quality_path):
        assert path.exists()

    package = json.loads(package_path.read_text(encoding="utf-8"))
    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
    validation = json.loads(validation_path.read_text(encoding="utf-8"))
    source_audit = json.loads(source_audit_path.read_text(encoding="utf-8"))
    coverage = json.loads(coverage_path.read_text(encoding="utf-8"))
    text_quality = json.loads(text_quality_path.read_text(encoding="utf-8"))

    expected_digest = _payload_sha256(package)
    assert source_audit["source_package_content_sha256"] == expected_digest
    assert coverage["source_package_content_sha256"] == expected_digest
    assert text_quality["source_package_content_sha256"] == expected_digest

    assert package["ticker_metadata"]["ticker"]["value"] == "ANF"
    assert package["company_profile"]["company_name"]["value"] == "Abercrombie & Fitch Co."
    assert len(package["quarterly_financials"]["rows"]) >= 8
    assert package["quarterly_financials"]["rows"][-1]["revenue"]["status"] == "populated"
    assert len(package["annual_financials"]["rows"]) >= 3
    assert package["debt_liquidity"]["cash"]["status"] == "populated"
    assert package["debt_liquidity"]["total_debt"]["status"] == "missing_source"
    assert package["debt_liquidity"]["total_debt"]["value"] is None
    assert package["debt_liquidity"]["net_leverage"]["status"] == "missing_source"
    assert package["debt_liquidity"]["total_liquidity"]["value"] == 1209.086
    assert package["debt_liquidity"]["as_of_date"]["value"] == "2026-01-31"
    assert len(package["normalized_guidance"]["items"]) >= 8
    assert len(package["segments"]["items"]) >= 6
    assert len(package["operating_drivers"]["items"]) >= 4
    assert len(package["quarter_notes"]["items"]) == 6
    assert package["source_coverage"]["sources"]
    latest_guidance = [
        item
        for item in package["normalized_guidance"]["items"]
        if item["publication_date"] == "2026-03-04"
    ]
    assert latest_guidance
    assert all(item["source_date"] == "2026-01-31" for item in latest_guidance)
    assert all(item["stated_in_period"] == "2025-Q4" for item in latest_guidance)
    assert all(item["display_role"] in {"current_primary", "current_secondary"} for item in latest_guidance)

    assert mapping["ticker"] == "ANF"
    assert isinstance(mapping["gaps"], list)
    assert validation["ticker"] == "ANF"
    assert not [issue for issue in validation["issues"] if issue["severity"] in {"P0", "P1"}]
    assert validate_normalized_company_data(package) == []

    audited_sections = {row["section"] for row in source_audit["sections"]}
    assert {
        "ticker_metadata",
        "company_profile",
        "quarterly_financials",
        "annual_financials",
        "debt_liquidity",
        "capital_returns",
        "normalized_guidance",
        "segments",
        "operating_drivers",
        "quarter_notes",
        "investment_case",
        "source_coverage",
        "mapping_gaps",
        "manual_review_flags",
    } <= audited_sections

    binding_ids = {row["binding_id"] for row in coverage["bindings"]}
    expected_binding_ids = {
        entry["binding_id"]
        for entry in json.loads((ROOT / "docs" / "workbook_binding_map.json").read_text(encoding="utf-8"))["bindings"]
    }
    assert binding_ids == expected_binding_ids
    assert any(row["would_write_useful_output"] for row in coverage["bindings"])
    assert text_quality["non_clean_visible_count"] == 0
    assert not (output_dir / "ANF_model.xlsx").exists()


def test_checked_in_anf_audits_match_current_authoritative_package_digest() -> None:
    package = json.loads((ANF_STRESS_DIR / "ANF_normalized_data_package.json").read_text(encoding="utf-8"))
    expected = _payload_sha256(package)

    for name in (
        "anf_normalized_package_source_audit.json",
        "anf_binding_coverage_audit.json",
        "anf_normalized_text_quality_audit.json",
    ):
        audit = json.loads((ROOT / "docs" / name).read_text(encoding="utf-8"))
        assert audit["source_package_content_sha256"] == expected, name


def test_anf_binding_coverage_reports_row_schema_capacity(tmp_path: Path) -> None:
    paths = build_anf_shadow_outputs(
        data_root=DATA_ROOT,
        workbook_path=ANF_WORKBOOK,
        output_dir=tmp_path / "ANF_new_ticker_engine",
        docs_dir=tmp_path / "docs",
    )

    coverage = json.loads(paths["binding_coverage_json"].read_text(encoding="utf-8"))
    rows = {row["binding_id"]: row for row in coverage["bindings"]}

    guidance = rows["pp_annual_guidance_rows"]
    quarter_notes = rows["qn_quarter_note_rows"]
    operating_drivers = rows["od_watchlist_rows"]
    qa_rows = rows["qa_checks_mapping_gap_rows"]
    quarterly_segments = rows["bs_segment_quarterly_rows"]
    annual_segments = rows["bs_segment_annual_rows"]

    assert set(guidance["row_schema_columns"]) >= {
        "metric",
        "initial_guide",
        "q1_update",
        "q2_update",
        "q3_update",
        "q4_update",
        "actual",
        "status",
        "notes_source",
    }
    assert guidance["number_of_values_available"] >= 1
    assert set(quarter_notes["row_schema_columns"]) >= {
        "theme",
        "quarter",
        "metric",
        "commentary",
        "model_implication",
        "source",
    }
    assert set(operating_drivers["row_schema_columns"]) >= {"topic", "current_read", "source", "why_it_matters"}
    assert set(qa_rows["row_schema_columns"]) >= {
        "rule_id",
        "status",
        "unique_issue_count",
        "occurrence_count",
        "blocking_count",
        "actionable_count",
        "affected_sections",
        "interpretation",
        "detail_ref",
    }
    latest_revenue = rows["summary_latest_revenue"]
    assert latest_revenue["planning_state"] == "active"
    assert latest_revenue["number_of_rows_planner_eligible"] == 1
    assert latest_revenue["structured_exclusion_count"] >= 1
    assert any(reason.startswith("row_selector_pick_excluded:") for reason in latest_revenue["structured_exclusion_reasons"])
    assert quarterly_segments["number_of_values_planner_eligible"] > 0
    assert quarterly_segments["planner_planned_write_count"] == 73
    assert quarterly_segments["would_write_useful_output"] is True
    assert annual_segments["number_of_values_planner_eligible"] > 0
    assert annual_segments["number_of_values_planner_eligible"] == 21
    assert annual_segments["planner_planned_write_count"] == 27
    assert annual_segments["would_write_useful_output"] is True
    assert coverage["planner_status"] == "PASS"

    inactive = next(row for row in coverage["bindings"] if row["planning_state"] != "active")
    assert inactive["would_write_useful_output"] is False
    assert "planning_state" in inactive["blank_reason"]
    assert coverage["binding_contract_content_sha256"]


def test_binding_coverage_independently_reproduces_and_rejects_cached_plan_tampering() -> None:
    package = json.loads(
        (ANF_STRESS_DIR / "ANF_normalized_data_package.json").read_text(encoding="utf-8")
    )
    bindings = json.loads((ROOT / "docs" / "workbook_binding_map.json").read_text(encoding="utf-8"))
    manifest = json.loads((ROOT / "docs" / "standard_template_shell_manifest.json").read_text(encoding="utf-8"))
    shell_identity = verify_shell_identity(
        ROOT / "templates" / "standard_stock_model_template.xlsx",
        manifest=manifest,
        binding_payload=bindings,
    )
    _plan, verified_plan = reproduce_binding_plan_snapshot(
        package,
        binding_payload=bindings,
        manifest=manifest,
        shell_path=ROOT / "templates" / "standard_stock_model_template.xlsx",
        shell_identity_report=shell_identity,
    )

    with pytest.raises(BindingPlanReproductionError, match="differs"):
        build_binding_coverage_audit(
            package,
            bindings,
            cached_plan={
                "status": "PASS",
                "planning_completed": True,
                "planned_writes": [{"binding_id": "fabricated", "target_sheet": "SUMMARY", "target_cell": "A3"}],
            },
            manifest=manifest,
            shell_path=ROOT / "templates" / "standard_stock_model_template.xlsx",
        )

    fabricated_plan = BindingPlan(ticker="ANF")
    fabricated_plan.planning_completed = True
    fabricated_plan.shell_identity_report = {"status": "PASS"}
    with pytest.raises(BindingPlanReproductionError, match="differs"):
        build_binding_coverage_audit(
            package,
            bindings,
            cached_plan=fabricated_plan,
            manifest=manifest,
            shell_path=ROOT / "templates" / "standard_stock_model_template.xlsx",
        )

    tampered_payload = verified_plan.plan_payload
    tampered_payload["planned_writes"][0]["target_sheet"] = "SUMMARY"
    tampered_payload["planned_writes"][0]["target_cell"] = "A3"
    tampered_payload["planned_writes"][0]["normalized_path"] = "fabricated.write"
    tampered_payload["planned_writes"][0]["value"] = "fabricated"
    tampered_payload["planned_writes"][0]["source_ref"] = "fabricated"
    tampered_token = copy.copy(verified_plan)
    object.__setattr__(
        tampered_token,
        "_plan_payload_json",
        json.dumps(tampered_payload, sort_keys=True, separators=(",", ":")),
    )
    object.__setattr__(
        tampered_token,
        "_consistency_json",
        json.dumps({"caller_recomputed_digest": _payload_sha256(tampered_payload)}, sort_keys=True),
    )
    with pytest.raises(BindingPlanReproductionError, match="differs"):
        build_binding_coverage_audit(
            package,
            bindings,
            cached_plan=tampered_token,
            manifest=manifest,
            shell_path=ROOT / "templates" / "standard_stock_model_template.xlsx",
        )

    changed_package = copy.deepcopy(package)
    changed_package["company_profile"]["business_description"]["value"] = "Changed after verification"
    with pytest.raises(BindingPlanReproductionError, match="differs"):
        build_binding_coverage_audit(
            changed_package,
            bindings,
            cached_plan=verified_plan,
            manifest=manifest,
            shell_path=ROOT / "templates" / "standard_stock_model_template.xlsx",
        )

    coverage = build_binding_coverage_audit(
        package,
        bindings,
        cached_plan=verified_plan,
        manifest=manifest,
        shell_path=ROOT / "templates" / "standard_stock_model_template.xlsx",
    )
    assert coverage["planner_status"] == "PASS"
    assert coverage["planner_total_write_count"] == len(verified_plan.plan_payload["planned_writes"])
    assert not any(
        write["normalized_path"] == "fabricated.write"
        for write in verified_plan.plan_payload["planned_writes"]
    )


def test_anf_shadow_package_demotes_noisy_visible_text(tmp_path: Path) -> None:
    paths = build_anf_shadow_outputs(
        data_root=DATA_ROOT,
        workbook_path=ANF_WORKBOOK,
        output_dir=tmp_path / "ANF_new_ticker_engine",
        docs_dir=tmp_path / "docs",
    )
    package = json.loads(paths["package"].read_text(encoding="utf-8"))
    text_quality = json.loads(paths["text_quality_json"].read_text(encoding="utf-8"))

    visible_blob = "\n".join(_visible_text_values(package))
    assert not re.search(r"compensation|governance|director|board|officer|restricted stock", visible_blob, re.I)
    assert not re.search(r"forward-looking|safe harbor|risk factors|trade policies or arrangements", visible_blob, re.I)
    assert "Gross profit divided by reported net sales" not in visible_blob
    assert "Operating income divided by reported net sales" not in visible_blob
    assert "REPORTS THIRD QUARTER" not in visible_blob
    assert not re.search(r"[-–]\s*$|\b(and|of|the|to|from|with)\s*$", visible_blob, re.I | re.M)

    demotions = package["source_coverage"].get("text_quality_demotions", [])
    assert demotions
    assert any(flag["rule_id"] == "text_quality_demoted" for flag in package["manual_review_flags"])
    assert text_quality["demotion_summary"]["total_demoted"] == len(demotions)
    assert build_normalized_text_quality_audit(package)["non_clean_visible_count"] == 0
    markdown = paths["text_quality_md"].read_text(encoding="utf-8")
    omitted = max(0, len(demotions) - 120)
    assert (f"{omitted} additional demotions omitted" in markdown) == bool(omitted)


def test_annual_aggregation_never_converts_missing_components_to_zero_or_partial_sum() -> None:
    missing = _annual_component_field(
        [("2025-Q1", 10.0), ("2025-Q2", None), ("2025-Q3", 30.0), ("2025-Q4", 40.0)],
        metric="adjusted_ebitda",
        period="2025-FY",
        source_ref="fixture:annual-components",
    )

    assert missing["value"] is None
    assert missing["status"] == "missing_source"
    assert missing["missing_inputs"] == ["2025-Q2"]

    package = build_anf_normalized_package(data_root=DATA_ROOT, workbook_path=ANF_WORKBOOK)
    annuals = {row["period"]: row for row in package["annual_financials"]["rows"]}
    for period in ("2020-FY", "2021-FY", "2022-FY"):
        field = annuals[period]["adjusted_ebitda"]
        assert field["value"] is None
        assert field["status"] == "missing_source"
        assert len(field["missing_inputs"]) == 4
        assert not any(value in (0, 0.0) for value in (field["value"],))
    reviews = [
        row
        for row in package["manual_review_flags"]
        if row.get("rule_id") == "legacy_adapter_annual_component_missing"
    ]
    assert {row["affected_period"] for row in reviews} >= {"2020-FY", "2021-FY", "2022-FY"}


def test_ttm_requires_exact_four_consecutive_source_backed_quarters() -> None:
    def rows(periods: list[str], *, units: dict[str, str] | None = None) -> list[dict]:
        return [
            {
                "period": period,
                "revenue": {
                    "value": float(index + 1),
                    "status": "populated",
                    "unit": (units or {}).get(period, "$m"),
                    "source_ref": f"fixture:{period}:{index}",
                },
            }
            for index, period in enumerate(periods)
        ]

    for periods in (["2025-Q4"], ["2025-Q3", "2025-Q4"], ["2025-Q2", "2025-Q3", "2025-Q4"]):
        result = _ttm_component_field(rows(list(periods)), metric="revenue", source_ref="fixture")
        assert result["status"] == "missing_source"
        assert result["value"] is None
        assert any(issue["reason"] == "quarter_count_not_four" for issue in result["component_issues"])

    non_consecutive = _ttm_component_field(
        rows(["2025-Q1", "2025-Q2", "2025-Q4", "2026-Q1"]),
        metric="revenue",
        source_ref="fixture",
    )
    assert non_consecutive["status"] == "manual_review_required"
    assert any(issue["reason"] == "quarters_not_consecutive" for issue in non_consecutive["component_issues"])

    duplicate = _ttm_component_field(
        rows(["2025-Q1", "2025-Q1", "2025-Q2", "2025-Q3"]),
        metric="revenue",
        source_ref="fixture",
        expected_end_period="2025-Q4",
    )
    assert duplicate["status"] == "manual_review_required"
    assert duplicate["duplicate_quarters"] == ["2025-Q1"]
    assert "2025-Q4" in duplicate["missing_quarters"]

    mismatched_unit = _ttm_component_field(
        rows(
            ["2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"],
            units={"2025-Q3": "%"},
        ),
        metric="revenue",
        source_ref="fixture",
    )
    assert mismatched_unit["status"] == "manual_review_required"
    assert any(issue["reason"] == "incompatible_unit" for issue in mismatched_unit["component_issues"])

    missing_unit_rows = rows(["2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"])
    missing_unit_rows[3]["revenue"].pop("unit")
    missing_unit = _ttm_component_field(
        missing_unit_rows,
        metric="revenue",
        source_ref="fixture",
    )
    assert missing_unit["status"] == "manual_review_required"
    assert missing_unit["value"] is None
    assert any(issue["reason"] == "incompatible_unit" for issue in missing_unit["component_issues"])

    for invalid_unit in ("   ", "bananas"):
        invalid_rows = rows(
            ["2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"],
            units={"2025-Q4": invalid_unit},
        )
        invalid_result = _ttm_component_field(
            invalid_rows,
            metric="revenue",
            source_ref="fixture",
        )
        assert invalid_result["status"] == "manual_review_required"
        assert invalid_result["value"] is None
        assert any(issue["reason"] == "incompatible_unit" for issue in invalid_result["component_issues"])

    mismatched_dimensions = rows(["2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"])
    mismatched_dimensions[0]["revenue"]["dimension"] = "geography"
    mismatched_dimensions[0]["revenue"]["member"] = "Region A"
    dimension_result = _ttm_component_field(
        mismatched_dimensions,
        metric="revenue",
        source_ref="fixture",
    )
    assert dimension_result["status"] == "manual_review_required"
    assert any(issue["reason"] == "incompatible_dimensions" for issue in dimension_result["component_issues"])

    valid_fiscal_year = _ttm_component_field(
        rows(["2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"]),
        metric="revenue",
        source_ref="fixture",
    )
    assert valid_fiscal_year["status"] == "populated"
    assert valid_fiscal_year["value"] == 10.0
    assert not any(
        issue["reason"] == "incompatible_unit"
        for issue in valid_fiscal_year.get("component_issues", [])
    )

    valid_rolling = _ttm_component_field(
        rows(["2025-Q2", "2025-Q3", "2025-Q4", "2026-Q1"]),
        metric="revenue",
        source_ref="fixture",
    )
    assert valid_rolling["status"] == "populated"
    assert valid_rolling["value"] == 10.0


def test_annual_aggregation_requires_exact_compatible_q1_q4_coverage() -> None:
    source = "fixture:annual-components"

    duplicate_q1 = _annual_component_field(
        [("2025-Q1", 1.0), ("2025-Q1", 2.0), ("2025-Q2", 3.0), ("2025-Q3", 4.0)],
        metric="revenue",
        period="2025-FY",
        source_ref=source,
    )
    assert duplicate_q1["status"] == "manual_review_required"
    assert any(row["reason"] == "duplicate_quarter" for row in duplicate_q1["component_issues"])
    assert "2025-Q4" in duplicate_q1["missing_inputs"]

    wrong_year = _annual_component_field(
        [
            {"label": "2025-Q1", "fiscal_year": 2024, "fiscal_quarter": 1, "value": 1.0, "unit": "$m", "status": "populated", "source_ref": source},
            *[
                {"label": f"2025-Q{quarter}", "fiscal_year": 2025, "fiscal_quarter": quarter, "value": float(quarter), "unit": "$m", "status": "populated", "source_ref": source}
                for quarter in (2, 3, 4)
            ],
        ],
        metric="revenue",
        period="2025-FY",
        source_ref=source,
    )
    assert any(row["reason"] == "mismatched_fiscal_year" for row in wrong_year["component_issues"])

    wrong_unit = _annual_component_field(
        [
            {"label": f"2025-Q{quarter}", "fiscal_year": 2025, "fiscal_quarter": quarter, "value": float(quarter), "unit": "%" if quarter == 3 else "$m", "status": "populated", "source_ref": source}
            for quarter in (1, 2, 3, 4)
        ],
        metric="revenue",
        period="2025-FY",
        source_ref=source,
    )
    assert any(row["reason"] == "mismatched_unit" for row in wrong_unit["component_issues"])

    valid = _annual_component_field(
        [(f"2025-Q{quarter}", float(quarter)) for quarter in (1, 2, 3, 4)],
        metric="revenue",
        period="2025-FY",
        source_ref=source,
    )
    assert valid["status"] == "populated"
    assert valid["value"] == 10.0


def test_missing_revenue_quarter_remains_an_explicit_annual_gap(monkeypatch) -> None:
    monkeypatch.setattr(
        "scripts.build_anf_shadow_normalized_package._read_legacy_valuation_series",
        lambda _path, _row: {},
    )
    rows = [
        {
            "fiscal_year": 2025,
            "fiscal_quarter": quarter,
            "fiscal_label": f"2025-Q{quarter}",
            "quarter": f"2025-0{quarter}-28",
            "revenue": None if quarter == 2 else 1_000_000.0 * quarter,
            "gross_profit": 500_000.0,
            "op_income": 200_000.0,
            "ebitda": 250_000.0,
            "net_income": 100_000.0,
            "cfo": 150_000.0,
            "capex": 50_000.0,
        }
        for quarter in (1, 2, 3, 4)
    ]

    incomplete_candidates: list[dict] = []
    annuals = _build_annual_financial_rows(
        rows,
        Path("legacy_fixture.xlsx"),
        incomplete_candidates=incomplete_candidates,
    )

    assert len(annuals) == 1
    assert annuals[0]["period"] == "2025-FY"
    assert annuals[0]["revenue"]["status"] == "missing_source"
    assert annuals[0]["revenue"]["value"] is None
    assert annuals[0]["operating_income"]["status"] == "populated"
    assert incomplete_candidates == []


def test_year_without_q4_becomes_an_explicit_incomplete_candidate(monkeypatch) -> None:
    monkeypatch.setattr(
        "scripts.build_anf_shadow_normalized_package._read_legacy_valuation_series",
        lambda _path, _row: {},
    )
    rows = [
        {
            "fiscal_year": 2024,
            "fiscal_quarter": quarter,
            "fiscal_label": f"2024-Q{quarter}",
            "quarter": f"2024-0{quarter}-28",
            "revenue": 1_000_000.0 * quarter,
        }
        for quarter in (1, 2, 3)
    ]
    incomplete_candidates: list[dict] = []

    annuals = _build_annual_financial_rows(
        rows,
        Path("legacy_fixture.xlsx"),
        incomplete_candidates=incomplete_candidates,
    )
    reviews = _annual_incomplete_candidate_reviews(incomplete_candidates)

    assert annuals == []
    assert incomplete_candidates == [
        {
            "period": "2024-FY",
            "status": "missing_source",
            "present_quarters": ["Q1", "Q2", "Q3"],
                "missing_quarters": ["Q4"],
                "source_refs": [
                    "legacy_fixture.xlsx!History_Q",
            ],
            "reason": "Annual aggregation requires exactly one source-backed Q1-Q4 component; missing Q4.",
        }
    ]
    assert reviews[0]["rule_id"] == "legacy_adapter_annual_fiscal_year_incomplete"
    assert reviews[0]["adapter_metadata"]["missing_quarters"] == ["Q4"]


def test_q4_only_year_remains_visible_as_missing_annual_row(monkeypatch) -> None:
    monkeypatch.setattr(
        "scripts.build_anf_shadow_normalized_package._read_legacy_valuation_series",
        lambda _path, _row: {},
    )
    rows = [
        {
            "fiscal_year": 2024,
            "fiscal_quarter": 4,
            "fiscal_label": "2024-Q4",
            "quarter": "2025-01-31",
            "revenue": 4_000_000.0,
        }
    ]
    incomplete_candidates: list[dict] = []

    annuals = _build_annual_financial_rows(
        rows,
        Path("legacy_fixture.xlsx"),
        incomplete_candidates=incomplete_candidates,
    )

    assert len(annuals) == 1
    assert annuals[0]["period"] == "2024-FY"
    assert annuals[0]["revenue"]["status"] == "missing_source"
    assert annuals[0]["revenue"]["missing_inputs"] == ["2024-Q1", "2024-Q2", "2024-Q3"]
    assert incomplete_candidates == []


def test_all_reliable_visible_legacy_segment_business_keys_are_normalized() -> None:
    package = build_anf_normalized_package(data_root=DATA_ROOT, workbook_path=ANF_WORKBOOK)
    normalized = {
        (item["dimension"], item["member"], item["period"], item["revenue"]["value"])
        for item in package["segments"]["items"]
        if item.get("source") == "legacy_visible_segment_oracle"
    }

    wb = load_workbook(ANF_WORKBOOK, read_only=True, data_only=True)
    try:
        ws = wb["BS_Segments"]
        expected = set()
        for period_type, header_row, member_rows, columns in (
            ("quarterly", 7, (61, 62, 63, 65, 66, 67), range(2, 14)),
            ("annual", 70, (72, 73, 74), range(2, 10)),
        ):
            for row_number in member_rows:
                member = str(ws.cell(row_number, 1).value or "")
                dimension = "geography" if member in {"Americas", "EMEA", "APAC"} else "brand" if member in {"Hollister", "Abercrombie"} else "total_company"
                for column in columns:
                    value = ws.cell(row_number, column).value
                    raw_period = ws.cell(header_row, column).value
                    if raw_period in (None, "") or not isinstance(value, (int, float)) or isinstance(value, bool):
                        continue
                    period = f"{int(raw_period)}-FY" if period_type == "annual" else str(raw_period)
                    expected.add((dimension, member, period, float(value)))
    finally:
        wb.close()

    assert len(expected) == 52
    assert normalized == expected


def test_annual_eps_and_net_debt_remain_missing_without_independent_lineage() -> None:
    package = build_anf_normalized_package(data_root=DATA_ROOT, workbook_path=ANF_WORKBOOK)

    for row in package["annual_financials"]["rows"]:
        assert row["eps"]["status"] == "missing_source"
        assert row["eps"]["value"] is None
        assert row["diluted_shares"]["status"] == "missing_source"
        assert row["diluted_shares"]["value"] is None
        if row["q4_diluted_shares"]["status"] == "populated":
            assert "retained separately for audit" in row["diluted_shares"]["reason"]

    assert package["valuation_inputs"]["net_debt"]["status"] == "missing_source"
    assert package["valuation_inputs"]["net_debt"]["value"] is None
    assert "D198 was not treated as evidence" in package["valuation_inputs"]["net_debt"]["reason"]
    assert package["valuation_inputs"]["book_value_per_share"]["status"] == "missing_source"
    assert "point-in-time shares outstanding" in package["valuation_inputs"]["book_value_per_share"]["reason"]

    for section in ("quarterly_financials", "annual_financials"):
        for row in package[section]["rows"]:
            for field in row.values():
                if isinstance(field, dict) and field.get("status") == "populated":
                    assert field.get("source_ref"), (section, row["period"])

    latest = package["quarterly_financials"]["rows"][-1]
    assert latest["net_income"]["definition"] == "Net income attributable to common shareholders."
    assert latest["eps"]["definition"] == "GAAP diluted earnings per share for the fiscal period."
    assert latest["diluted_shares"]["definition"] == "Quarterly weighted-average diluted shares used for diluted EPS."


def test_missing_fail_zero_placeholders_are_missing_and_never_aggregated() -> None:
    package = build_anf_normalized_package(data_root=DATA_ROOT, workbook_path=ANF_WORKBOOK)
    quarterly = {row["period"]: row for row in package["quarterly_financials"]["rows"]}
    expected = {
        "total_debt": {"2024-Q2", "2024-Q3", "2024-Q4"},
        "debt_core": {"2024-Q2", "2024-Q3", "2024-Q4", "2025-Q2", "2025-Q3", "2025-Q4"},
        "interest_paid": {"2024-Q3", "2024-Q4", "2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"},
    }
    for metric, periods in expected.items():
        for period in periods:
            field = quarterly[period][metric]
            assert field["status"] == "missing_source", (metric, period, field)
            assert field["value"] is None
            assert "zero was treated as a missing placeholder" in field["reason"]
            assert "REPORT_" in field["source_ref"]

    # The report-level Missing/FAIL marker identifies zero placeholders. It
    # does not erase older non-zero legacy evidence on the same support row.
    assert quarterly["2024-Q1"]["total_debt"]["status"] == "populated"
    assert quarterly["2024-Q1"]["debt_core"]["status"] == "populated"
    assert quarterly["2024-Q2"]["interest_paid"]["status"] == "populated"

    reviews = [
        row
        for row in package["manual_review_flags"]
        if row.get("rule_id") == "legacy_adapter_unsupported_zero_placeholder"
    ]
    assert len(reviews) == 15
    assert {(row["row_key"].split("|", 1)[1], row["affected_period"]) for row in reviews} == {
        (metric, period)
        for metric, periods in expected.items()
        for period in periods
    }
    assert all(row["source_ref"] and row["suggested_action"] for row in reviews)

    history_keys = {
        (item["metric"], item["period"])
        for item in package["calculation_history"]["quarterly_items"]
    }
    assert not any(
        (metric, period) in history_keys
        for metric, periods in expected.items()
        for period in periods
    )

    annual = {row["period"]: row for row in package["annual_financials"]["rows"]}
    for period, metric in (
        ("2024-FY", "total_debt"),
        ("2024-FY", "debt_core"),
        ("2024-FY", "interest_paid"),
        ("2025-FY", "debt_core"),
        ("2025-FY", "interest_paid"),
    ):
        field = annual[period][metric]
        assert field["status"] in {"missing_source", "manual_review_required"}
        assert field["value"] is None
        assert any(
            issue.get("reason") == "unsupported_zero_placeholder"
            for issue in field.get("component_issues", [])
        )

    assert package["valuation_inputs"]["interest_paid_ttm"]["status"] != "populated"
    assert package["valuation_inputs"]["interest_paid_ttm"]["value"] is None


def test_latest_missing_fail_zero_cannot_bypass_debt_liquidity(monkeypatch) -> None:
    monkeypatch.setattr(
        "scripts.build_anf_shadow_normalized_package._read_legacy_valuation_series",
        lambda _path, _row: {},
    )
    history_rows = [
        {
            "quarter": "2025-02-01",
            "fiscal_year": 2024,
            "fiscal_quarter": 4,
            "fiscal_label": "2024-Q4",
            "revenue": 1_000_000_000.0,
            "cash": 500_000_000.0,
            "total_debt": 0.0,
        }
    ]
    placeholders = {
        ("total_debt", "2025-02-01"): {
            "metric": "total_debt",
            "line_item": "Total debt",
            "sheet": "REPORT_BS_Q",
            "period_end": "2025-02-01",
            "candidate_value": 0.0,
            "source_status": "Missing",
            "qa_status": "FAIL",
            "value_source_ref": "ANF_model.xlsx!REPORT_BS_Q!M5",
            "metadata_source_ref": "ANF_model.xlsx!REPORT_BS_Q!C5:D5",
        }
    }
    review_flags: list[dict] = []
    quarterly_rows = _build_quarterly_financial_rows(
        history_rows,
        Path("ANF_model.xlsx"),
        unsupported_zero_placeholders=placeholders,
        review_flags=review_flags,
    )

    debt_liquidity = _build_debt_liquidity(
        history_rows,
        quarterly_rows,
        [],
        [],
        Path("ANF_model.xlsx"),
        review_flags,
    )
    valuation_inputs = _build_valuation_inputs(
        quarterly_rows,
        debt_liquidity,
        Path("ANF_model.xlsx"),
        review_flags=review_flags,
    )

    assert quarterly_rows[0]["total_debt"]["status"] == "missing_source"
    assert "REPORT_BS_Q" in quarterly_rows[0]["total_debt"]["source_ref"]
    for field in (
        debt_liquidity["total_debt"],
        debt_liquidity["net_debt"],
        debt_liquidity["net_leverage"],
        valuation_inputs["net_debt"],
    ):
        assert field["status"] == "missing_source"
        assert field["value"] is None
    assert any(
        row.get("rule_id") == "legacy_adapter_unsupported_zero_placeholder"
        and row.get("row_key") == "2024-Q4|total_debt"
        for row in review_flags
    )


def test_all_legacy_annual_candidates_are_retained_before_visible_capacity_selection() -> None:
    package = build_anf_normalized_package(data_root=DATA_ROOT, workbook_path=ANF_WORKBOOK)
    rows = package["annual_financials"]["rows"]
    periods = [row["period"] for row in rows]

    assert periods == [f"{year}-FY" for year in range(2014, 2026)]
    assert {"2018-FY", "2019-FY"} <= set(periods)
    assert {"2014-FY", "2015-FY", "2016-FY", "2017-FY"} <= set(periods)

    for period in ("2018-FY", "2019-FY"):
        row = next(item for item in rows if item["period"] == period)
        assert row["revenue"]["status"] == "populated"
        assert row["revenue"]["source_ref"]
        assert row["cost_of_goods_sold"]["status"] == "populated"

    plan = json.loads(ANF_PLAN.read_text(encoding="utf-8"))
    axis = plan["period_axes"]["bs_annual_financial_periods"]
    assert list(axis["period_to_column"]) == [f"{year}-FY" for year in range(2018, 2026)]
    report = next(row for row in plan["bindings"] if row["binding_id"] == "financial_fact_disposition_audit")
    older = [
        row
        for row in report["skipped_rows"]
        if row["section"] == "annual_financials"
        and row["normalized_path"].split(".")[2] in {"0", "1", "2", "3"}
    ]
    assert older
    assert all(row["disposition"] in {"audit_only", "formula_owned"} for row in older)
    assert all(row["reason"] for row in older)


def test_calculation_history_covers_visible_axis_plus_seven_prior_quarters() -> None:
    package = build_anf_normalized_package(data_root=DATA_ROOT, workbook_path=ANF_WORKBOOK)
    visible_periods = [row["period"] for row in package["quarterly_financials"]["rows"]]
    revenue_ordinals = {
        int(item["period_ordinal"])
        for item in package["calculation_history"]["quarterly_items"]
        if item["metric"] == "revenue"
    }

    def ordinal(period: str) -> int:
        return int(period[:4]) * 4 + int(period[-1]) - 1

    expected = set(range(min(map(ordinal, visible_periods)) - 7, max(map(ordinal, visible_periods)) + 1))
    assert expected <= revenue_ordinals


def test_every_populated_anf_financial_fact_has_an_explicit_disposition() -> None:
    package = json.loads(ANF_PACKAGE.read_text(encoding="utf-8"))
    plan = json.loads(ANF_PLAN.read_text(encoding="utf-8"))
    report = next(row for row in plan["bindings"] if row["binding_id"] == "financial_fact_disposition_audit")
    assert report["unresolved_fact_count"] == 0
    assert report["populated_fact_count"] == report["planned_fact_count"] + report["explicit_disposition_count"]
    assert len(report["skipped_rows"]) == report["explicit_disposition_count"]
    assert all(row["disposition"] in {"formula_owned", "audit_only", "explicitly_excluded"} for row in report["skipped_rows"])
    assert all(row["reason"] and row["normalized_path"] and row["source_ref"] for row in report["skipped_rows"])

    disposition_counts = Counter(row["disposition"] for row in report["skipped_rows"])
    assert disposition_counts.get("explicitly_excluded", 0) == 0
    assert disposition_counts.get("formula_owned", 0) > 0
    assert disposition_counts.get("audit_only", 0) > 0

    for row in report["skipped_rows"]:
        if row["field"] == "free_cash_flow":
            assert row["disposition"] == "formula_owned"
            assert "row_selector" not in row["reason"]
        if row["field"] == "total_debt":
            assert row["disposition"] == "audit_only"
            assert "row_selector" not in row["reason"]

    # Preserve the reviewed pre-expansion reconciliation after removing the 20
    # quarterly/annual debt and cash-interest placeholders. The final package
    # additionally contains older annual candidates and newly classified
    # COGS/tax/D&A/operating-margin evidence.
    added_quarterly_fields = {
        "cost_of_goods_sold",
        "income_taxes_paid",
        "depreciation_amortization",
        "operating_margin",
    }
    added_annual_fields = {
        "cost_of_goods_sold",
        "income_taxes_paid",
        "depreciation_amortization",
        "debt_current",
    }
    baseline_paths: list[str] = []
    for section, rows, offset, added_fields in (
        ("quarterly_financials", package["quarterly_financials"]["rows"], 0, added_quarterly_fields),
        (
            "annual_financials",
            package["annual_financials"]["rows"][-6:],
            len(package["annual_financials"]["rows"]) - 6,
            added_annual_fields,
        ),
    ):
        for row_index, item in enumerate(rows, start=offset):
            for field, node in item.items():
                if field in added_fields or not isinstance(node, dict):
                    continue
                if node.get("status") == "populated" and node.get("value") not in (None, ""):
                    baseline_paths.append(f"{section}.rows.{row_index}.{field}")

    planned_paths = {write["normalized_path"] for write in plan["planned_writes"]}
    dispositions_by_path = {row["normalized_path"]: row["disposition"] for row in report["skipped_rows"]}
    baseline_reconciliation = Counter(
        "planned"
        if path in planned_paths
        else dispositions_by_path.get(path, "unexplained")
        for path in baseline_paths
    )
    assert len(baseline_paths) == 523
    assert baseline_reconciliation == Counter(
        {"planned": 397, "formula_owned": 18, "audit_only": 108}
    )

    writes = plan["planned_writes"]
    assert not any(write["normalized_path"] == "valuation_inputs.net_debt" for write in writes)
    assert not any(write["target_sheet"] == "Valuation" and write["target_cell"] == "D198" for write in writes)


def _visible_text_values(package: dict) -> list[str]:
    values: list[str] = []
    for item in package.get("quarter_notes", {}).get("items", []):
        for key in ("note", "commentary", "model_implication", "valuation_implication"):
            values.append(_field_text(item.get(key)))
    for item in package.get("operating_drivers", {}).get("items", []):
        for key in ("driver", "current_read", "why_it_matters"):
            values.append(_field_text(item.get(key)))
    for item in package.get("segments", {}).get("items", []):
        values.append(_field_text(item.get("note")))
    for item in package.get("normalized_guidance", {}).get("items", []):
        values.append(str(item.get("source_excerpt") or ""))
    return [value for value in values if value]


def _field_text(value) -> str:
    if isinstance(value, dict):
        return str(value.get("value") or "")
    return str(value or "")
