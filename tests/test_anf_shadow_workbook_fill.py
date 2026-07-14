from __future__ import annotations

from copy import deepcopy
import json
from pathlib import Path
import shutil
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, range_boundaries
import pytest

from pbi_xbrl.new_ticker_binding_planner import (
    BindingPlanReproductionError,
    reproduce_binding_plan,
)
from pbi_xbrl.standard_template_shell_identity import _planned_cell_values_equal, verify_post_fill_structural_identity
import scripts.fill_anf_shadow_workbook as shadow_workflow
from scripts.fill_anf_shadow_workbook import _atomic_promote_workbook, run_anf_shadow_workbook_fill
from scripts.validate_standard_template_shell import validate_shell


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "standard_stock_model_template.xlsx"
MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
DATA_ROOT = next(
    ancestor / "StockModelData"
    for ancestor in [ROOT, *ROOT.parents]
    if (ancestor / "StockModelData").exists()
)
PACKAGE = DATA_ROOT / "outputs" / "stress_tests" / "ANF_new_ticker_engine" / "ANF_normalized_data_package.json"
CACHED_PLAN = DATA_ROOT / "outputs" / "stress_tests" / "ANF_new_ticker_engine" / "ANF_binding_plan.json"
LEGACY_ANF = DATA_ROOT / "outputs" / "Excel stock models" / "ANF_model.xlsx"


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _strict_post_fill_report(workbook_path: Path, plan_path: Path) -> dict[str, Any]:
    return validate_shell(
        template_path=workbook_path,
        manifest_path=MANIFEST,
        binding_map_path=BINDING_MAP,
        allow_filled_values=True,
        approved_shell_path=TEMPLATE,
        approved_plan_path=plan_path,
        normalized_package_path=PACKAGE,
    )


@pytest.fixture(scope="module")
def anf_shadow_artifacts(tmp_path_factory: pytest.TempPathFactory) -> dict[str, Path]:
    output_dir = tmp_path_factory.mktemp("anf-shadow-e2e") / "ANF_new_ticker_engine"
    return run_anf_shadow_workbook_fill(
        package_path=PACKAGE,
        output_dir=output_dir,
        legacy_workbook_path=LEGACY_ANF,
        cached_plan_path=CACHED_PLAN,
    )


def _find_unplanned_owned_cell(workbook_path: Path, plan: dict[str, Any]) -> tuple[str, str]:
    binding_payload = _load_json(BINDING_MAP)
    planned = {
        (str(write["target_sheet"]), str(write["target_cell"]))
        for write in plan.get("planned_writes") or []
    }
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        for binding in binding_payload.get("bindings") or []:
            if not binding.get("writable") or binding.get("planning_state", "active") != "active":
                continue
            sheet = str(binding.get("sheet") or "").replace("{ticker}", "ANF")
            if sheet not in wb.sheetnames:
                continue
            target = str(binding.get("planner_target") or binding.get("target") or "")
            min_col, min_row, max_col, max_row = range_boundaries(target)
            target_columns = [
                column_index_from_string(str(column["target_column"]))
                for column in binding.get("target_columns") or []
                if isinstance(column, dict) and column.get("target_column")
            ]
            columns = target_columns or list(range(min_col, max_col + 1))
            for row in range(min_row, max_row + 1):
                for column in columns:
                    cell = wb[sheet].cell(row, column)
                    if (
                        (sheet, cell.coordinate) not in planned
                        and not isinstance(cell, MergedCell)
                        and not (isinstance(cell.value, str) and cell.value.startswith("="))
                    ):
                        return sheet, cell.coordinate
    finally:
        wb.close()
    raise AssertionError("Expected at least one exact writable cell not selected by the ANF plan.")


def _candidate_workbooks(output_dir: Path) -> list[Path]:
    return list(output_dir.glob(".ANF_shadow_model.*.candidate.xlsx"))


def test_anf_shadow_fill_uses_reproduced_plan_and_strict_post_fill_validation(
    anf_shadow_artifacts: dict[str, Path],
) -> None:
    paths = anf_shadow_artifacts
    expected = {
        "workbook",
        "plan_json",
        "prefill_json",
        "prefill_txt",
        "postfill_json",
        "postfill_txt",
        "comparison_json",
        "comparison_txt",
    }
    assert expected <= set(paths)
    for key in expected:
        assert paths[key].exists(), key
    assert paths["workbook"].name == "ANF_shadow_model.xlsx"
    assert not (paths["workbook"].parent / "ANF_model.xlsx").exists()
    assert _candidate_workbooks(paths["workbook"].parent) == []

    package = _load_json(PACKAGE)
    manifest = _load_json(MANIFEST)
    binding_payload = _load_json(BINDING_MAP)
    plan = _load_json(paths["plan_json"])
    cached_plan = _load_json(CACHED_PLAN)
    reproduced = reproduce_binding_plan(
        package,
        manifest=manifest,
        binding_payload=binding_payload,
        shell_path=TEMPLATE,
        ticker_override="ANF",
        expected_plan=plan,
    )
    assert reproduced.to_dict() == plan == cached_plan

    prefill = _load_json(paths["prefill_json"])
    postfill = _load_json(paths["postfill_json"])
    comparison = _load_json(paths["comparison_json"])
    shell_report = _strict_post_fill_report(paths["workbook"], paths["plan_json"])

    assert prefill["status"] == "PASS"
    assert prefill["visible_rows_available"]["quarterly_financial_rows"] >= 8
    assert prefill["visible_rows_available"]["annual_financial_rows"] >= 3
    assert prefill["visible_rows_available"]["guidance_rows"] >= 5
    assert prefill["visible_rows_available"]["segment_rows"] >= 5
    assert prefill["visible_rows_available"]["operating_driver_visible_rows"] >= 5
    assert prefill["visible_rows_available"]["quarter_note_visible_rows"] >= 5
    assert prefill["demoted_rows"]["total_demoted"] > 0

    assert postfill["status"] == "PASS"
    assert postfill["strict_post_fill_validation"]["status"] == "PASS"
    assert postfill["strict_post_fill_validation"]["issue_count"] == 0
    assert postfill["approved_plan_status"] == "PASS"
    assert postfill["approved_plan_write_count"] == plan["planned_write_count"]
    assert postfill["layout_signature_unchanged"] is True
    assert postfill["formulas_unchanged"] is True
    assert postfill["non_writable_cells_unchanged"] is True
    assert shell_report["status"] == "PASS", shell_report["issues"][:10]

    business_writes = [
        write
        for write in plan["planned_writes"]
        if write["target_sheet"] not in {"QA_Log", "Needs_Review", "QA_Checks"}
    ]
    assert len({(write["target_sheet"], write["target_cell"]) for write in business_writes}) == len(business_writes)
    business_bindings = [write["binding_id"] for write in business_writes]
    assert business_bindings.count("valuation_period_headers") == 12
    assert business_bindings.count("valuation_revenue_series") == 12
    assert business_bindings.count("valuation_net_income_series") == 12
    assert business_bindings.count("bs_annual_financial_period_headers") == 6
    assert business_bindings.count("bs_annual_revenue_series") == 8
    by_target = {
        f"{write['target_sheet']}!{write['target_cell']}": write["value"]
        for write in plan["planned_writes"]
    }
    wb = load_workbook(paths["workbook"], data_only=False, read_only=False)
    try:
        assert "ANF_Investment_Case" in wb.sheetnames
        for write in plan["planned_writes"]:
            assert _planned_cell_values_equal(
                wb[write["target_sheet"]][write["target_cell"]].value,
                write["value"],
            )

        assert wb["Operating_Drivers"]["A6"].value == by_target["Operating_Drivers!A6"]
        assert wb["Operating_Drivers"]["B6"].value == by_target["Operating_Drivers!B6"]
        assert wb["Operating_Drivers"]["H6"].value == by_target["Operating_Drivers!H6"]
        assert wb["Quarter_Notes_UI"]["A9"].value == "Theme"
        assert wb["Quarter_Notes_UI"]["A10"].value == by_target["Quarter_Notes_UI!A10"]
        assert wb["Quarter_Notes_UI"]["C10"].value == by_target["Quarter_Notes_UI!C10"]
        assert wb["Promise_Progress_UI"]["A61"].value == by_target["Promise_Progress_UI!A61"]
        assert wb["Promise_Progress_UI"]["C61"].value == by_target["Promise_Progress_UI!C61"]
    finally:
        wb.close()

    assert comparison["summary"]["blocks_compared"] > 0
    assert comparison["summary"]["shadow_populated_blocks"] > 0
    assert comparison["top_binding_gaps_to_fix_next"] == []


def test_anf_strict_post_fill_requires_reproduction_inputs(anf_shadow_artifacts: dict[str, Path]) -> None:
    report = validate_shell(
        template_path=anf_shadow_artifacts["workbook"],
        manifest_path=MANIFEST,
        binding_map_path=BINDING_MAP,
        allow_filled_values=True,
        approved_shell_path=TEMPLATE,
    )

    assert report["status"] == "FAIL"
    assert "post_fill_reproduction_inputs_missing" in {issue["rule_id"] for issue in report["issues"]}


def test_anf_shadow_workflow_rejects_stale_cached_plan_before_workbook_copy(tmp_path: Path) -> None:
    stale = deepcopy(_load_json(CACHED_PLAN))
    stale["planned_writes"][0]["value"] = "STALE CACHED VALUE"
    stale_path = tmp_path / "stale-plan.json"
    stale_path.write_text(json.dumps(stale), encoding="utf-8")
    output_dir = tmp_path / "stale-plan-output"

    with pytest.raises(BindingPlanReproductionError, match="differs"):
        run_anf_shadow_workbook_fill(
            package_path=PACKAGE,
            output_dir=output_dir,
            legacy_workbook_path=LEGACY_ANF,
            cached_plan_path=stale_path,
        )

    assert not (output_dir / "ANF_shadow_model.xlsx").exists()


def test_anf_unchanged_shell_rejects_fabricated_plan() -> None:
    package = _load_json(PACKAGE)
    manifest = _load_json(MANIFEST)
    binding_payload = _load_json(BINDING_MAP)
    fabricated = deepcopy(_load_json(CACHED_PLAN))
    fabricated["planned_writes"][0].update(
        {
            "target_sheet": "SUMMARY",
            "target_cell": "A3",
            "value": "FORGED VALUE",
            "source_ref": "fabricated",
        }
    )
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        report = verify_post_fill_structural_identity(
            wb,
            approved_shell_path=TEMPLATE,
            manifest=manifest,
            binding_payload=binding_payload,
            approved_plan=fabricated,
            normalized_package=package,
        )
    finally:
        wb.close()

    assert report["changed_writable_cell_count"] == 0
    assert report["status"] == "FAIL"
    assert "post_fill_binding_plan_reproduction_failed" in {issue["rule_id"] for issue in report["issues"]}


def test_anf_strict_post_fill_rejects_non_writable_drift(
    anf_shadow_artifacts: dict[str, Path], tmp_path: Path
) -> None:
    drifted = tmp_path / "protected-drift.xlsx"
    shutil.copyfile(anf_shadow_artifacts["workbook"], drifted)
    wb = load_workbook(drifted, data_only=False, read_only=False)
    try:
        wb["SUMMARY"]["A1"] = "Unauthorized protected change"
        wb.save(drifted)
    finally:
        wb.close()

    report = _strict_post_fill_report(drifted, anf_shadow_artifacts["plan_json"])
    assert report["status"] == "FAIL"
    assert "post_fill_protected_cell_drift" in {issue["rule_id"] for issue in report["issues"]}


def test_anf_strict_post_fill_rejects_missing_planned_value(
    anf_shadow_artifacts: dict[str, Path], tmp_path: Path
) -> None:
    drifted = tmp_path / "missing-planned-value.xlsx"
    shutil.copyfile(anf_shadow_artifacts["workbook"], drifted)
    plan = _load_json(anf_shadow_artifacts["plan_json"])
    write = next(row for row in plan["planned_writes"] if row["target_sheet"] not in {"QA_Log", "Needs_Review", "QA_Checks"})
    wb = load_workbook(drifted, data_only=False, read_only=False)
    try:
        wb[write["target_sheet"]][write["target_cell"]] = None
        wb.save(drifted)
    finally:
        wb.close()

    report = _strict_post_fill_report(drifted, anf_shadow_artifacts["plan_json"])
    assert report["status"] == "FAIL"
    assert "post_fill_planned_value_mismatch" in {issue["rule_id"] for issue in report["issues"]}


def test_anf_strict_post_fill_rejects_unplanned_writable_value(
    anf_shadow_artifacts: dict[str, Path], tmp_path: Path
) -> None:
    plan = _load_json(anf_shadow_artifacts["plan_json"])
    sheet, coordinate = _find_unplanned_owned_cell(anf_shadow_artifacts["workbook"], plan)
    drifted = tmp_path / "unplanned-value.xlsx"
    shutil.copyfile(anf_shadow_artifacts["workbook"], drifted)
    wb = load_workbook(drifted, data_only=False, read_only=False)
    try:
        wb[sheet][coordinate] = "UNPLANNED VALUE"
        wb.save(drifted)
    finally:
        wb.close()

    report = _strict_post_fill_report(drifted, anf_shadow_artifacts["plan_json"])
    assert report["status"] == "FAIL"
    assert "post_fill_unplanned_value_change" in {issue["rule_id"] for issue in report["issues"]}


def test_atomic_shadow_promotion_replaces_only_the_final_path(tmp_path: Path) -> None:
    final_path = tmp_path / "ANF_shadow_model.xlsx"
    candidate_path = tmp_path / ".ANF_shadow_model.test.candidate.xlsx"
    final_path.write_bytes(b"previous-approved-output")
    candidate_path.write_bytes(b"strictly-validated-candidate")

    _atomic_promote_workbook(candidate_path, final_path)

    assert final_path.read_bytes() == b"strictly-validated-candidate"
    assert not candidate_path.exists()


@pytest.mark.parametrize("existing_output", [False, True])
def test_shadow_validation_failure_cleans_candidate_and_preserves_final_output(
    existing_output: bool,
    anf_shadow_artifacts: dict[str, Path],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_dir = tmp_path / ("existing-final" if existing_output else "no-final")
    output_dir.mkdir(parents=True)
    final_path = output_dir / "ANF_shadow_model.xlsx"
    if existing_output:
        shutil.copyfile(anf_shadow_artifacts["workbook"], final_path)
    before = final_path.read_bytes() if final_path.exists() else None

    def copy_shell_candidate(
        _package_path: Path,
        *,
        output_path: Path,
        template_path: Path,
        **_kwargs: Any,
    ) -> None:
        shutil.copyfile(template_path, output_path)

    monkeypatch.setattr(shadow_workflow, "fill_standard_template_from_package", copy_shell_candidate)
    monkeypatch.setattr(
        shadow_workflow,
        "validate_shell",
        lambda **_kwargs: {
            "status": "FAIL",
            "issue_count": 1,
            "issues": [{"rule_id": "forced_transaction_failure", "message": "test failure"}],
            "shell_identity": {},
        },
    )

    with pytest.raises(RuntimeError, match="strict post-fill validation failed"):
        shadow_workflow.run_anf_shadow_workbook_fill(
            package_path=PACKAGE,
            output_dir=output_dir,
            legacy_workbook_path=LEGACY_ANF,
            cached_plan_path=CACHED_PLAN,
        )

    assert _candidate_workbooks(output_dir) == []
    if existing_output:
        assert final_path.read_bytes() == before
    else:
        assert not final_path.exists()
