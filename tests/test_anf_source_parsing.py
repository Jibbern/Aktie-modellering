import datetime as dt

from openpyxl import Workbook
import pandas as pd
import pytest

from pbi_xbrl.pipeline_orchestration import (
    _apply_anf_company_overview_overrides,
    _build_anf_guidance_progress_rows,
    _build_anf_source_quarter_notes,
    _dedupe_local_non_gaap_segment_rows,
    _dedupe_slides_guidance_rows,
    _normalize_anf_guidance_rows,
    _parse_anf_adjusted_metrics_from_lines,
    _parse_anf_guidance_rows_from_lines,
    _parse_anf_quarterly_history_retail_driver_rows,
    _parse_anf_retail_text_driver_rows_from_lines,
    _parse_anf_sales_mix_tables_from_html,
    _parse_anf_statement_values_from_lines,
    _sanitize_anf_adjusted_metric_units,
)
from pbi_xbrl.pipeline_types import PipelineConfig


Q4_2025_LINES = [
    "Thirteen Weeks Ended Thirteen Weeks Ended",
    "January 31, 2026 Sales February 1, 2025 Sales",
    "Net sales $ 1,669,802 100.0 % $ 1,584,917 100.0 %",
    "Cost of sales, exclusive of depreciation and amortization 676,491 40.5 % 610,907 38.5 %",
    "Operating income 235,931 14.1 % 256,064 16.2 %",
    "Net income attributable to A&F $ 172,130 10.3 % $ 187,226 11.8 %",
    "Net income per share attributable to A&F",
    "Diluted $ 3.68 $ 3.57",
    "Weighted-average shares outstanding:",
    "Diluted 46,837 52,461",
    "Schedule of Non-GAAP Financial Measures",
    "Fifty-Two Weeks Ended January 31, 2026",
    "GAAP (1) Net Sales Excluded item (2) non-GAAP Net Sales",
    "Litigation settlement $ (38,574) $ (38,574) $ —",
    "Operating income 699,143 13.3 % 38,574 660,569 12.5 %",
    "Net income attributable to A&F 506,921 9.6 % 28,882 478,039 9.1 %",
    "Net income per diluted share attributable to A&F $ 10.46 $ 0.60 $ 9.86",
    "Diluted weighted-average shares outstanding 48,476 48,476",
    "Reconciliation of EBITDA and Adjusted EBITDA",
    "Thirteen Weeks Ended January 31, 2026 and Thirteen Weeks Ended February 1, 2025",
    "EBITDA (1) $ 276,386 16.6 $ 293,227 18.5",
    "Reconciliation of EBITDA and Adjusted EBITDA",
    "Fifty-Two Weeks Ended January 31, 2026 and Fifty-Two Weeks Ended February 1, 2025",
    "EBITDA (1) $ 854,164 16.2 $ 894,593 18.1",
    "Adjustments to EBITDA",
    "Litigation settlement (38,574) (0.7) — —",
    "Adjusted EBITDA (1) $ 815,590 15.5 $ 894,593 18.1",
]


def test_anf_statement_parser_captures_q4_shares_eps_and_ebitda() -> None:
    values = _parse_anf_statement_values_from_lines(Q4_2025_LINES, scale=1000.0)

    assert values["revenue"] == pytest.approx(1_669_802_000.0)
    assert values["gross_profit"] == pytest.approx(993_311_000.0)
    assert values["op_income"] == pytest.approx(235_931_000.0)
    assert values["net_income"] == pytest.approx(172_130_000.0)
    assert values["eps_diluted"] == pytest.approx(3.68)
    assert values["shares_diluted"] == pytest.approx(46_837_000.0)
    assert values["ebitda"] == pytest.approx(276_386_000.0)


def test_anf_adjusted_parser_keeps_quarter_and_annual_metrics_separate() -> None:
    rows = _parse_anf_adjusted_metrics_from_lines(
        Q4_2025_LINES,
        quarter_end=dt.date(2026, 1, 31),
        scale=1000.0,
        source_doc="ANF_Q4_2025_earnings_presentation_financial_schedules.pdf",
        source="slides",
    )

    by_period = {row["period_type"]: row for row in rows}
    assert by_period["quarter"]["adj_ebitda"] == pytest.approx(276_386_000.0)
    assert by_period["annual"]["adj_ebitda"] == pytest.approx(815_590_000.0)
    assert by_period["annual"]["adj_ebit"] == pytest.approx(660_569_000.0)
    assert by_period["annual"]["adj_eps"] == pytest.approx(9.86)
    assert "favorable settlement" in by_period["annual"]["source_snippet"].lower()


def test_anf_guidance_parser_captures_q1_and_fy2026_outlook() -> None:
    lines = [
        "Fourth Quarter Fiscal 2025 Results",
        "Net sales increased in the fourth quarter of fiscal 2025.",
        "Fiscal 2026 First Quarter and Full Year Outlook",
        "For fiscal 2026, the company expects:",
        "First Quarter Outlook (1) Full Year Outlook (1)",
        "Net sales Growth In The Range of 1% to 3% Growth In The Range of 3% to 5%",
        "Operating margin Around 7.0% In the Range of 12.0% to 12.5%",
        "Net income per diluted share (3) (4) In The Range of $1.20 to $1.30 In The Range of $10.20 to $11.00",
        "Share repurchases (4) At least $100 million Around $450 million",
        "Diluted weighted average shares (3) (4) Around 46 million Around 45 million",
        "Capital expenditures In The Range of $200 to $225 million",
        "Real estate activity (5)(all approximate) ~30 Net Store Openings",
        "Real estate activity (5)(all approximate) 55 Openings, 25 Closures",
        "Real estate activity (5)(all approximate) 70 Remodels and Right-Sizes",
        "the outlook assumes a year-over-year tariff impact as a percentage of net sales of approximately 290 basis points for the first quarter and 70 basis points for the full year.",
    ]

    rows = _parse_anf_guidance_rows_from_lines(lines, dt.date(2026, 1, 31))
    labels = {(row["period_label"], row["metric_hint"]): row for row in rows}

    assert labels[("Q1 FY2026", "Revenue")]["numbers"] == "1%, 3%"
    assert labels[("FY2026", "Revenue")]["numbers"] == "3%, 5%"
    assert labels[("FY2026", "Capex")]["numbers"] == "$200 million, $225 million"
    assert labels[("Q1 FY2026", "Share repurchases")]["numbers"] == "at least $100 million"
    assert labels[("FY2026", "Real estate activity")]["numbers"] == "55 openings, 25 closures; 70 remodels/right-sizes"
    assert labels[("Q1 FY2026", "Tariffs")]["numbers"] == "290 bps"


def test_anf_guidance_parser_handles_future_updated_outlook_period_labels() -> None:
    lines = [
        "Fiscal 2026 Second Quarter and Full Year Updated Outlook",
        "Second Quarter Outlook Full Year Outlook",
        "Net sales Growth in the range of 2% to 4% Growth in the range of 4% to 6%",
        "Operating margin Around 10.0% In the range of 12.5% to 13.0%",
        "Net income per diluted share In the range of $2.00 to $2.20 In the range of $11.00 to $12.00",
        "Share repurchases At least $125 million Around $500 million",
        "Diluted weighted average shares Around 45 million Around 44 million",
        "Capital expenditures In the range of $225 to $250 million",
        "The outlook assumes a tariff impact of approximately 200 basis points for the second quarter and 60 basis points for the full year.",
    ]

    rows = _parse_anf_guidance_rows_from_lines(lines, dt.date(2026, 5, 2))
    labels = {(row["period_label"], row["metric_hint"]): row for row in rows}

    assert labels[("Q2 FY2026", "Revenue")]["numbers"] == "2%, 4%"
    assert labels[("FY2026", "Revenue")]["numbers"] == "4%, 6%"
    assert labels[("Q2 FY2026", "Operating margin")]["numbers"] == "around 10.0%"
    assert labels[("FY2026", "Operating margin")]["numbers"] == "12.5%, 13.0%"
    assert labels[("Q2 FY2026", "Adj EPS")]["numbers"] == "$2.00, $2.20"
    assert labels[("FY2026", "Capex")]["numbers"] == "$225 million, $250 million"
    assert labels[("Q2 FY2026", "Tariffs")]["numbers"] == "200 bps"


def test_anf_summary_override_replaces_noisy_overview_and_uses_apac_mix() -> None:
    overview = {
        "what_it_does": "It operates through Corporate / Other.",
        "key_advantage": "Form of Retention Restricted Stock Unit Award Agreement under the plan.",
        "current_strategic_context": "Code of Business Conduct and corporate governance.",
        "revenue_streams": [{"name": "Americas", "pct": 0.92}, {"name": "EMEA", "pct": 0.08}],
    }
    segments = pd.DataFrame(
        [
            {"quarter": "2026-01-31", "segment": "Americas", "metric": "revenue", "value": 4_290_395_000.0, "period_type": "annual"},
            {"quarter": "2026-01-31", "segment": "EMEA", "metric": "revenue", "value": 818_140_000.0, "period_type": "annual"},
            {"quarter": "2026-01-31", "segment": "APAC", "metric": "revenue", "value": 157_757_000.0, "period_type": "annual"},
        ]
    )

    cleaned = _apply_anf_company_overview_overrides(overview, slides_segments=segments)

    assert "Corporate / Other" not in cleaned["what_it_does"]
    assert "restricted stock unit" not in cleaned["key_advantage"].lower()
    assert "Code of Business Conduct" not in cleaned["current_strategic_context"]
    stream_names = {row["name"] for row in cleaned["revenue_streams"]}
    assert stream_names == {"Americas", "EMEA", "APAC"}
    apac = next(row for row in cleaned["revenue_streams"] if row["name"] == "APAC")
    assert apac["pct"] == pytest.approx(0.03, abs=0.001)


def test_anf_adjusted_metric_sanitizer_normalizes_generic_sec_1000x_rows() -> None:
    adj_metrics = pd.DataFrame(
        [
            {
                "quarter": pd.Timestamp("2025-11-01"),
                "adj_ebit": 155_021_000_000.0,
                "adj_ebitda": 193_587_000_000.0,
                "source_snippet": "generic SEC parse",
            },
            {
                "quarter": pd.Timestamp("2026-01-31"),
                "period_type": "annual",
                "adj_ebit": 660_569_000.0,
                "adj_ebitda": 815_590_000.0,
                "source_snippet": "ANF financial schedule",
            },
        ]
    )
    adj_breakdown = pd.DataFrame(
        [
            {"quarter": pd.Timestamp("2025-11-01"), "label": "Tax effect", "value": 12_000_000_000.0},
            {"quarter": pd.Timestamp("2026-01-31"), "label": "Favorable legal settlement", "value": -38_574_000.0},
        ]
    )
    hist = pd.DataFrame(
        [
            {"quarter": pd.Timestamp("2025-11-01"), "revenue": 1_208_011_000.0},
            {"quarter": pd.Timestamp("2026-01-31"), "revenue": 1_669_802_000.0},
        ]
    )

    clean_metrics, clean_breakdown = _sanitize_anf_adjusted_metric_units(adj_metrics, adj_breakdown, hist)

    q3 = clean_metrics[clean_metrics["quarter"].eq(pd.Timestamp("2025-11-01"))].iloc[0]
    fy = clean_metrics[clean_metrics["quarter"].eq(pd.Timestamp("2026-01-31"))].iloc[0]
    assert q3["adj_ebit"] == pytest.approx(155_021_000.0)
    assert q3["adj_ebitda"] == pytest.approx(193_587_000.0)
    assert fy["adj_ebit"] == pytest.approx(660_569_000.0)
    assert "unit-normalized" in q3["source_snippet"]
    bd_q3 = clean_breakdown[clean_breakdown["quarter"].eq(pd.Timestamp("2025-11-01"))].iloc[0]
    assert bd_q3["value"] == pytest.approx(12_000_000.0)


def test_anf_source_notes_skip_zero_revenue_and_impossible_margin(tmp_path) -> None:
    earnings_dir = tmp_path / "earnings_presentation"
    earnings_dir.mkdir()
    (earnings_dir / "ANF_Q1_2024_earnings_presentation.txt").write_text(
        "\n".join(
            [
                "Condensed Consolidated Statements of Operations",
                "Thirteen Weeks Ended May 4, 2024",
                "Net sales 100.0% 100.0%",
                "Cost of sales, exclusive of depreciation and amortization 0.0% 0.0%",
                "Gross profit 100.0% 100.0%",
                "Operating income 100.0% 100.0%",
            ]
        ),
        encoding="utf-8",
    )
    hist = pd.DataFrame(
        [
            {
                "quarter": pd.Timestamp("2024-05-04"),
                "revenue": 1_020_730_000.0,
                "gross_profit": 672_765_000.0,
                "op_income": 130_000_000.0,
            }
        ]
    )

    notes = _build_anf_source_quarter_notes(
        hist=hist,
        base_dir=tmp_path,
        config=PipelineConfig(cache_dir=tmp_path / "cache"),
        max_quarters=1,
    )
    note_blob = " | ".join(notes.get("note", pd.Series(dtype=str)).astype(str).tolist())

    assert "$0.0m" not in note_blob
    assert "gross margin 100.0%" not in note_blob


def test_anf_segment_dedupe_rejects_tiny_revenue_rows() -> None:
    rows = pd.DataFrame(
        [
            {"quarter": "2024-11-02", "segment": "APAC", "metric": "revenue", "value": 26.0, "unit": "USD", "period_type": "quarter"},
            {"quarter": "2024-11-02", "segment": "Americas", "metric": "revenue", "value": 14.0, "unit": "USD", "period_type": "quarter"},
            {"quarter": "2024-11-02", "segment": "EMEA", "metric": "revenue", "value": 6.0, "unit": "USD", "period_type": "quarter"},
            {"quarter": "2024-11-02", "segment": "APAC", "metric": "revenue", "value": 40_925_000.0, "unit": "USD", "period_type": "quarter"},
        ]
    )

    out = _dedupe_local_non_gaap_segment_rows(rows)

    assert set(out["segment"]) == {"APAC"}
    assert out.iloc[0]["value"] == pytest.approx(40_925_000.0)


def test_slides_guidance_dedupes_by_period_metric_numbers_and_source_doc() -> None:
    rows = pd.DataFrame(
        [
            {
                "quarter": "2026-01-31",
                "period_label": "FY2026",
                "metric_hint": "Revenue",
                "numbers": "3%, 5%",
                "doc": "ANF_Q4_2025_earnings_release.htm",
                "source": "earnings_release",
                "line": "FY2026 net sales growth 3% to 5%",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "FY2026",
                "metric_hint": "Revenue",
                "numbers": "3%, 5%",
                "doc": "ANF_Q4_2025_earnings_release.htm",
                "source": "earnings_release",
                "line": "FY2026 net sales growth 3% to 5%",
            },
        ]
    )

    out = _dedupe_slides_guidance_rows(rows)

    assert len(out) == 1


def test_anf_guidance_normalizer_rejects_stale_prior_period_rows_from_later_sources() -> None:
    raw = pd.DataFrame(
        [
            {
                "quarter": "2026-01-31",
                "period_label": "Q1 FY2025",
                "metric_hint": "Adj EPS",
                "numbers": "$1.20, $1.30",
                "line": "First quarter outlook for adjusted EPS in the range of $1.20 to $1.30",
                "source": "earnings_release",
                "doc": "ANF_Q4_2025_earnings_release.htm",
                "unit": "$",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "FY2025",
                "metric_hint": "Revenue",
                "numbers": "6%",
                "line": "Fiscal 2025 outlook net sales growth at least 6%",
                "source": "earnings_release",
                "doc": "ANF_Q4_2025_earnings_release.htm",
                "unit": "%",
            },
            {
                "quarter": "2026-01-12",
                "period_label": "FY2025",
                "metric_hint": "Revenue",
                "numbers": "6%",
                "line": "Business update: currently expects fiscal 2025 net sales growth at least 6%",
                "source": "press_release",
                "doc": "ANF_January_2026_business_update.htm",
                "unit": "%",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "FY2025",
                "metric_hint": "Revenue",
                "numbers": "6%, 7%",
                "line": "2025 year Revenue 6%, 7%",
                "source": "press_release",
                "doc": "ANF_January_2026_business_update.htm",
                "unit": "%",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "Q1 FY2026",
                "metric_hint": "Adj EPS",
                "numbers": "$1.20, $1.30",
                "line": "First quarter outlook for adjusted EPS in the range of $1.20 to $1.30",
                "source": "earnings_release",
                "doc": "ANF_Q4_2025_earnings_release.htm",
                "unit": "$",
            },
        ]
    )

    out = _normalize_anf_guidance_rows(raw)
    kept = {(row["doc"], row["period_label"], row["metric_hint"]) for row in out.to_dict("records")}

    assert ("ANF_Q4_2025_earnings_release.htm", "Q1 FY2025", "Adj EPS") not in kept
    assert ("ANF_Q4_2025_earnings_release.htm", "FY2025", "Revenue") not in kept
    assert ("ANF_January_2026_business_update.htm", "FY2025", "Revenue") in kept
    assert ("ANF_Q4_2025_earnings_release.htm", "Q1 FY2026", "Adj EPS") in kept
    jan_rows = out[out["doc"].eq("ANF_January_2026_business_update.htm") & out["period_label"].eq("FY2025") & out["metric_hint"].eq("Revenue")]
    assert len(jan_rows) == 1
    assert jan_rows.iloc[0]["stated_in_label"] == "Jan 2026 pre-release update"


def test_anf_brand_family_revenue_and_comps_from_release_tables(tmp_path) -> None:
    html_path = tmp_path / "ANF_Q4_2025_earnings_release.htm"
    html_path.write_text(
        """
        <table>
          <tr><th colspan="5">Fourth Quarter Fiscal 2025</th></tr>
          <tr><td>Net sales by brand family</td><td>2025</td><td>2024</td><td>% Change</td><td>Comparable Sales</td></tr>
          <tr><td>Abercrombie</td><td>806,502</td><td>775,858</td><td>4%</td><td>(1)%</td></tr>
          <tr><td>Hollister</td><td>863,300</td><td>809,059</td><td>6%</td><td>3%</td></tr>
          <tr><td>Total company</td><td>1,669,802</td><td>1,584,917</td><td>5%</td><td>1%</td></tr>
        </table>
        <table>
          <tr><th colspan="5">Full Year Fiscal 2025</th></tr>
          <tr><td>Net sales by segment</td><td>2025</td><td>2024</td><td>% Change</td><td>Comparable Sales</td></tr>
          <tr><td>Americas</td><td>4,290,395</td><td>4,054,726</td><td>6%</td><td>4%</td></tr>
          <tr><td>EMEA</td><td>818,140</td><td>816,066</td><td>0%</td><td>0%</td></tr>
          <tr><td>APAC</td><td>157,757</td><td>154,769</td><td>2%</td><td>(3)%</td></tr>
          <tr><td>Total company</td><td>5,266,292</td><td>5,025,561</td><td>5%</td><td>3%</td></tr>
          <tr><td>Net sales by brand family</td><td>2025</td><td>2024</td><td>% Change</td><td>Comparable Sales</td></tr>
          <tr><td>Abercrombie</td><td>2,523,662</td><td>2,546,143</td><td>(1)%</td><td>(7)%</td></tr>
          <tr><td>Hollister</td><td>2,742,630</td><td>2,479,418</td><td>15%</td><td>13%</td></tr>
        </table>
        """,
        encoding="utf-8",
    )

    rows = _parse_anf_sales_mix_tables_from_html(html_path, dt.date(2026, 1, 31))
    key = {
        (str(row["period_type"]), str(row["segment"]), str(row["metric"])): row
        for row in rows.to_dict("records")
    }

    assert key[("quarter", "Abercrombie", "revenue")]["value"] == pytest.approx(806_502_000.0)
    assert key[("quarter", "Abercrombie", "comparable_sales")]["value"] == pytest.approx(-0.01)
    assert key[("quarter", "Hollister", "comparable_sales")]["value"] == pytest.approx(0.03)
    assert key[("annual", "Abercrombie", "revenue")]["value"] == pytest.approx(2_523_662_000.0)
    assert key[("annual", "Hollister", "revenue")]["value"] == pytest.approx(2_742_630_000.0)
    assert key[("annual", "Hollister", "net_sales_growth")]["value"] == pytest.approx(0.15)
    assert key[("annual", "Hollister", "comparable_sales")]["value"] == pytest.approx(0.13)
    assert key[("annual", "APAC", "revenue")]["value"] == pytest.approx(157_757_000.0)


def test_anf_quarterly_history_extracts_fy2025_comps_and_store_count(tmp_path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Historical Comparable Sales"
    for col in range(2, 7):
        ws.cell(row=4, column=col, value="Fiscal 2025")
    for col, label in enumerate(["Q1", "Q2", "Q3", "Q4", "2025"], start=2):
        ws.cell(row=5, column=col, value=label)
    rows = [
        ("Total comparable sales", 0.04, 0.03, 0.03, 0.01, 0.03),
        ("Abercrombie comparable sales", -0.10, -0.11, -0.07, -0.01, -0.07),
        ("Hollister comparable sales", 0.23, 0.19, 0.15, 0.03, 0.13),
    ]
    for ridx, row in enumerate(rows, start=6):
        for cidx, value in enumerate(row, start=1):
            ws.cell(row=ridx, column=cidx, value=value)

    store = wb.create_sheet("Store Count")
    store.cell(row=5, column=1, value="Company-owned stores")
    store.cell(row=5, column=2, value="Total Company")
    store.cell(row=5, column=3, value="Abercrombie")
    store.cell(row=5, column=4, value="Hollister")
    for ridx, row in enumerate(
        [
            ("February 1, 2025", 789, 278, 511),
            ("New stores", 62, 36, 26),
            ("Permanently closed", -22, -8, -14),
            ("January 31, 2026", 829, 306, 523),
            ("Franchise stores", 60, 37, 23),
            ("Total including franchise", 889, 343, 546),
        ],
        start=6,
    ):
        for cidx, value in enumerate(row, start=1):
            store.cell(row=ridx, column=cidx, value=value)

    path = tmp_path / "ANF_Q4_2025_earnings_presentation_quarterly_history.xlsx"
    wb.save(path)
    fiscal_periods = {
        (2025, 1): dt.date(2025, 5, 3),
        (2025, 2): dt.date(2025, 8, 2),
        (2025, 3): dt.date(2025, 11, 1),
        (2025, 4): dt.date(2026, 1, 31),
    }

    out = _parse_anf_quarterly_history_retail_driver_rows(path, fiscal_periods=fiscal_periods)
    key = {
        (
            pd.Timestamp(row["quarter"]).date(),
            str(row["segment"]),
            str(row["metric"]),
            str(row["period_type"]),
        ): row
        for row in out.to_dict("records")
    }

    assert key[(dt.date(2025, 5, 3), "Total Company", "comparable_sales", "quarter")]["value"] == pytest.approx(0.04)
    assert key[(dt.date(2025, 8, 2), "Abercrombie", "comparable_sales", "quarter")]["value"] == pytest.approx(-0.11)
    assert key[(dt.date(2025, 11, 1), "Hollister", "comparable_sales", "quarter")]["value"] == pytest.approx(0.15)
    assert key[(dt.date(2026, 1, 31), "Hollister", "comparable_sales", "quarter")]["value"] == pytest.approx(0.03)
    assert key[(dt.date(2026, 1, 31), "Total Company", "store_count_beginning", "annual")]["value"] == pytest.approx(789)
    assert key[(dt.date(2026, 1, 31), "Total Company", "new_stores", "annual")]["value"] == pytest.approx(62)
    assert key[(dt.date(2026, 1, 31), "Total Company", "closed_stores", "annual")]["value"] == pytest.approx(22)
    assert key[(dt.date(2026, 1, 31), "Total Company", "store_count_end", "quarter")]["value"] == pytest.approx(829)
    assert key[(dt.date(2026, 1, 31), "Total Company", "franchise_stores", "quarter")]["value"] == pytest.approx(60)
    assert key[(dt.date(2026, 1, 31), "Total Company", "total_stores_including_franchise", "quarter")]["value"] == pytest.approx(889)


def test_anf_retail_text_driver_extracts_digital_inventory_margin_and_buybacks() -> None:
    lines = [
        "Both brands had record fourth quarter net sales; Abercrombie returned to growth in Q4 and Hollister delivered its 11th consecutive quarter of growth.",
        "For the year, 44% of total sales were digital, with Hollister around 31% and Abercrombie around 59%. Our platforms had more than 1 billion visits and the omnichannel customer is our most valuable customer.",
        "We delivered 120 new store experiences, including 62 new stores, 11 right sizes and 47 remodels. We also closed 22 stores and ended the year with 829 stores, 523 Hollister and 306 Abercrombie.",
        "Inventory cost was up 5%, including roughly 3 points from tariffs. Units were up 5%, including roughly 3 points from strategic receipts ahead of the ERP go-live, or up roughly 2% excluding ERP.",
        "Q1 includes a tariff headwind of approximately 290 basis points, or $30 million, and a freight tailwind of approximately 160 basis points. ERP creates a 1 to 2 percentage point sales headwind and over 100 basis points of operating margin headwind. Marketing spend is up about 50 basis points as a percentage of sales.",
        "For fiscal 2026, tariffs are expected to be about 70 basis points, or $40 million incremental, with mitigation through sourcing, supplier negotiation, product costing and selective pricing. We expect slight AUR expansion.",
        "Full year share repurchases were $450 million, or 5.4 million shares, representing 11% of shares outstanding at February 1, 2025, with $850 million remaining authorization.",
    ]

    out = _parse_anf_retail_text_driver_rows_from_lines(
        lines,
        quarter_end=dt.date(2026, 1, 31),
        source_doc="ANF_Q4_2025_transcript.txt",
        source_type="transcript",
    )
    metrics = {str(row["metric"]) for row in out.to_dict("records")}

    assert "digital_sales_mix" in metrics
    assert "digital_visits" in metrics
    assert "store_count_end" in metrics
    assert "inventory_cost_growth" in metrics
    assert "inventory_unit_growth_ex_erp" in metrics
    assert "q1_fy2026_tariff_headwind_bps" in metrics
    assert "q1_fy2026_freight_tailwind_bps" in metrics
    assert "fy2026_tariff_headwind_bps" in metrics
    assert "share_repurchases" in metrics
    avg_price = out.loc[out["metric"].eq("average_buyback_price"), "value"].iloc[0]
    assert avg_price == pytest.approx(83.33, abs=0.01)


def test_anf_buyback_source_note_does_not_mislabel_tariff_cost_as_full_year_buybacks(tmp_path) -> None:
    transcript_dir = tmp_path / "earnings_transcripts"
    transcript_dir.mkdir()
    (transcript_dir / "ANF_Q4_2025_transcript.txt").write_text(
        "\n".join(
            [
                "Full year share repurchases were $450 million, or 5.4 million shares, representing 11% of shares outstanding at February 1, 2025, with $850 million remaining authorization; tariff impact around $90 million cost pressure.",
            ]
        ),
        encoding="utf-8",
    )
    hist = pd.DataFrame(
        [
            {
                "quarter": pd.Timestamp("2026-01-31"),
                "revenue": 1_669_802_000.0,
                "gross_profit": 993_311_000.0,
                "op_income": 235_931_000.0,
            }
        ]
    )

    notes = _build_anf_source_quarter_notes(
        hist=hist,
        base_dir=tmp_path,
        config=PipelineConfig(cache_dir=tmp_path / "cache"),
        max_quarters=1,
    )
    note_blob = " | ".join(notes.get("note", pd.Series(dtype=str)).astype(str).tolist())

    assert "2025 year repurchases were about $90m" not in note_blob
    assert "FY2025 repurchases were about $90m" not in note_blob
    assert "2025 year repurchases were about $450m" in note_blob
    assert "5.4m shares" in note_blob


def test_anf_guidance_normalization_excludes_actuals_headers_and_wrong_units() -> None:
    raw = pd.DataFrame(
        [
            {
                "quarter": "2026-01-31",
                "period_label": "FY2026",
                "metric_hint": "Revenue",
                "numbers": "3%, 5%",
                "unit": "%",
                "line": "Full Year Outlook net sales growth in the range of 3% to 5%",
                "source": "earnings_release",
                "doc": "ANF_Q4_2025_earnings_release.htm",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "FY2026",
                "metric_hint": "Revenue",
                "numbers": "$90 million",
                "unit": "$m",
                "line": "tariff cost impact around $90 million for fiscal 2025",
                "source": "earnings_release",
                "doc": "ANF_Q2_2025_earnings_release.htm",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "FY2025",
                "metric_hint": "Revenue",
                "numbers": "5%",
                "unit": "%",
                "line": "Net sales increased 5% in actual fourth quarter results",
                "source": "earnings_release",
                "doc": "ANF_Q4_2025_earnings_release.htm",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "FY2026",
                "metric_hint": "Operating margin",
                "numbers": "12.0%, 12.5%",
                "unit": "%",
                "line": "ANF_Q4_2025_earnings_release.htm",
                "source": "earnings_release",
                "doc": "ANF_Q4_2025_earnings_release.htm",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "FY2026",
                "metric_hint": "Adj EPS",
                "numbers": "$10.20, $11.00",
                "unit": "$",
                "line": "the following risk factors and safe harbor statements apply",
                "source": "annual_reports",
                "doc": "ANF_2025_10K.htm",
            },
        ]
    )

    out = _normalize_anf_guidance_rows(raw)

    assert len(out) == 1
    row = out.iloc[0]
    assert row["metric_hint"] == "Revenue"
    assert row["numbers"] == "3%, 5%"
    assert row["unit"] == "%"


def test_anf_guidance_progress_tracks_fy2025_revisions_and_actuals() -> None:
    guidance = pd.DataFrame(
        [
            {"quarter": "2025-02-01", "period_label": "FY2025", "metric_hint": "Revenue", "numbers": "3%, 5%", "unit": "%", "line": "Fiscal 2025 outlook sales growth 3% to 5%"},
            {"quarter": "2025-05-03", "period_label": "FY2025", "metric_hint": "Revenue", "numbers": "3%, 6%", "unit": "%", "line": "Updated fiscal 2025 outlook sales growth 3% to 6%"},
            {"quarter": "2025-08-02", "period_label": "FY2025", "metric_hint": "Revenue", "numbers": "5%, 7%", "unit": "%", "line": "Updated fiscal 2025 outlook sales growth 5% to 7%"},
            {"quarter": "2025-11-01", "period_label": "FY2025", "metric_hint": "Adj EPS", "numbers": "$10.20, $10.50", "unit": "$", "line": "Updated fiscal 2025 EPS $10.20 to $10.50"},
            {"quarter": "2026-01-31", "period_label": "FY2026", "metric_hint": "Revenue", "numbers": "3%, 5%", "unit": "%", "line": "Fiscal 2026 outlook sales growth 3% to 5%"},
        ]
    )
    hist = pd.DataFrame(
        [
            {"quarter": "2025-02-01", "revenue": 5_025_561_000.0, "op_income": 837_882_000.0, "eps_diluted": 9.00, "capex": 0.0},
            {"quarter": "2026-01-31", "revenue": 5_266_292_000.0, "op_income": 699_143_000.0, "eps_diluted": 10.46, "capex": 240_774_000.0},
        ]
    )
    adj = pd.DataFrame([{"quarter": "2026-01-31", "period_type": "annual", "adj_eps": 9.86, "adj_ebit": 660_569_000.0}])

    out = _build_anf_guidance_progress_rows(guidance, hist=hist, adj_metrics=adj)

    fy2025_revenue = out[out["promise_id"].astype(str).str.contains("FY2025:Revenue")]
    fy2026_revenue = out[out["promise_id"].astype(str).str.contains("FY2026:Revenue")]
    assert len(fy2025_revenue) >= 3
    assert fy2025_revenue.iloc[-1]["status"] in {"resolved_pass", "resolved_watch"}
    assert "actual" in str(fy2025_revenue.iloc[-1]["rationale"]).lower()
    assert len(fy2026_revenue) == 1
    assert fy2026_revenue.iloc[0]["status"] == "open"


def test_anf_guidance_normalizer_reclassifies_annual_eps_from_q1_prefix() -> None:
    raw = pd.DataFrame(
        [
            {
                "quarter": "2025-05-03",
                "period_label": "Q1 FY2025",
                "metric_hint": "Adj EPS",
                "numbers": "$9.50, $10.50",
                "unit": "$",
                "line": "Fiscal 2025 Full Year Outlook adjusted net income per diluted share $9.50 to $10.50",
                "source": "earnings_release",
                "doc": "ANF_Q1_2025_earnings_release.pdf",
            },
            {
                "quarter": "2025-05-03",
                "period_label": "Q1 FY2025",
                "metric_hint": "Revenue",
                "numbers": "approximately 90",
                "unit": "",
                "line": "Fiscal 2025 outlook approximately 90",
                "source": "earnings_release",
                "doc": "ANF_Q1_2025_earnings_release.pdf",
            },
        ]
    )

    out = _normalize_anf_guidance_rows(raw)

    assert len(out) == 1
    assert out.iloc[0]["metric_hint"] == "Adj EPS"
    assert out.iloc[0]["period_label"] == "FY2025"
    assert out.iloc[0]["period_type"] == "annual"
    assert out.iloc[0]["horizon_label"] == "2025 year"
    assert out.iloc[0]["stated_in_label"] == "Q1 2025"


def test_anf_source_notes_include_brand_specific_latest_quarter_note(tmp_path) -> None:
    transcript_dir = tmp_path / "earnings_transcripts"
    transcript_dir.mkdir()
    (transcript_dir / "ANF_Q4_2025_transcript.txt").write_text(
        "\n".join(
            [
                "Both brands had record fourth quarter net sales.",
                "Abercrombie returned to growth in Q4 and Hollister delivered its 11th consecutive quarter of growth.",
                "For the year, 44% of total sales were digital, with more than 1 billion visits.",
            ]
        ),
        encoding="utf-8",
    )
    hist = pd.DataFrame(
        [
            {
                "quarter": pd.Timestamp("2026-01-31"),
                "revenue": 1_669_802_000.0,
                "gross_profit": 993_311_000.0,
                "op_income": 235_931_000.0,
            }
        ]
    )

    notes = _build_anf_source_quarter_notes(
        hist=hist,
        base_dir=tmp_path,
        config=PipelineConfig(cache_dir=tmp_path / "cache"),
        max_quarters=1,
    )
    note_blob = " | ".join(notes.get("note", pd.Series(dtype=str)).astype(str).tolist())

    assert "Abercrombie returned to growth" in note_blob
    assert "Hollister" in note_blob
