from __future__ import annotations

import json
from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import pytest
from openpyxl import load_workbook

from pbi_xbrl.json_schema_validation import load_json_strict
from pbi_xbrl.longitudinal_memory.summary_bs_workbook_materialization import (
    _sheet_part_map,
    sha256_file,
)
from pbi_xbrl.longitudinal_memory.valuation_capital_product_cleanup import (
    EXPECTED_EXPANDED_PREVIEW_SHA256,
    EXPECTED_INVESTOR_PRODUCT_DIGEST,
    HIDDEN_LINEAGE_RANGE,
    VISIBLE_CAPITAL_RANGE,
    build_valuation_capital_product_cleanup_plan,
    materialize_valuation_capital_product_cleanup,
)


ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = Path(r"C:\Users\Jibbe\Aktier\StockModelData")
BASE = (
    DATA_ROOT
    / "audit"
    / "capital_allocation_return_product_expansion_2026-08-16"
    / "ANF_capital_allocation_return_expansion_preview_a.xlsx"
)
PACKAGE = (
    DATA_ROOT
    / "outputs"
    / "stress_tests"
    / "ANF_new_ticker_engine"
    / "ANF_normalized_data_package.json"
)
BS_PRODUCT = ROOT / "tests" / "fixtures" / "summary_bs" / "anf_bs_segment_product.v1.json"
BS_SHADOW = ROOT / "tests" / "fixtures" / "summary_bs" / "anf_bs_segment_shadow.v1.json"


def _require_inputs() -> None:
    for path in (BASE, PACKAGE, BS_PRODUCT, BS_SHADOW):
        if not path.is_file():
            pytest.skip(f"Required accepted input is unavailable: {path}")


@pytest.fixture(scope="session")
def plan():
    _require_inputs()
    return build_valuation_capital_product_cleanup_plan(
        package=load_json_strict(PACKAGE),
        source_package_path=PACKAGE,
        balance_sheet_product=load_json_strict(BS_PRODUCT),
        balance_sheet_product_path=BS_PRODUCT,
        balance_sheet_shadow=load_json_strict(BS_SHADOW),
        balance_sheet_shadow_path=BS_SHADOW,
        base_workbook=BASE,
    )


@pytest.fixture(scope="session")
def materialized(tmp_path_factory: pytest.TempPathFactory, plan):
    output = tmp_path_factory.mktemp("valuation_capital_cleanup") / "preview.xlsx"
    result = materialize_valuation_capital_product_cleanup(
        plan=plan,
        base_workbook=BASE,
        output_workbook=output,
    )
    return output, result


def _defined_names(path: Path) -> dict[str, str]:
    namespace = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    names = root.find("m:definedNames", namespace)
    return {
        f"{node.attrib['name']}|{node.attrib.get('localSheetId', '')}": node.text or ""
        for node in (() if names is None else names)
    }


def _formula_map(path: Path) -> dict[str, str]:
    workbook = load_workbook(path, data_only=False)
    try:
        return {
            f"{sheet.title}!{cell.coordinate}": str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        }
    finally:
        workbook.close()


def test_accepted_base_and_product_identity(plan) -> None:
    assert sha256_file(BASE) == EXPECTED_EXPANDED_PREVIEW_SHA256
    assert plan.base_workbook_sha256 == EXPECTED_EXPANDED_PREVIEW_SHA256
    assert plan.investor_product["product_digest"] == EXPECTED_INVESTOR_PRODUCT_DIGEST
    assert VISIBLE_CAPITAL_RANGE == "A126:M166"
    assert HIDDEN_LINEAGE_RANGE == "A270:A297"


def test_plan_replay_is_deterministic(plan) -> None:
    repeat = build_valuation_capital_product_cleanup_plan(
        package=load_json_strict(PACKAGE),
        source_package_path=PACKAGE,
        balance_sheet_product=load_json_strict(BS_PRODUCT),
        balance_sheet_product_path=BS_PRODUCT,
        balance_sheet_shadow=load_json_strict(BS_SHADOW),
        balance_sheet_shadow_path=BS_SHADOW,
        base_workbook=BASE,
    )
    assert repeat.to_dict() == plan.to_dict()


def test_capital_product_relocated_with_required_hierarchy(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert sheet["A126"].value == "Capital Allocation"
        assert sheet["A127"].value == "Summary"
        assert sheet["A133"].value == "Annual History"
        assert sheet["A140"].value == "Capital Return"
        assert sheet["A141"].value == "Summary"
        assert sheet["A151"].value == "Quarterly History"
        assert sheet["A159"].value == "Annual History"
        assert [sheet.cell(139, column).value for column in range(1, 14)] == [None] * 13
        assert sheet.row_dimensions[139].height == pytest.approx(8.1)
    finally:
        workbook.close()


def test_section_lettering_and_combined_title_are_retired(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        values = [
            cell.value
            for row in workbook["Valuation"].iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        ]
        assert "Capital Allocation & Capital Return" not in values
        assert not any(value.startswith(("A. ", "B. ", "C. ", "D. ", "E. ")) for value in values)
        assert values.count("Capital Allocation") == 1
        assert values.count("Capital Return") == 1
    finally:
        workbook.close()


def test_quarterly_history_uses_natural_a_to_m_fit(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert [sheet.cell(152, column).value for column in range(1, 14)] == [
            "Metric",
            "Q2'23",
            "Q3'23",
            "Q4'23",
            "Q1'24",
            "Q2'24",
            "Q3'24",
            "Q4'24",
            "Q1'25",
            "Q2'25",
            "Q3'25",
            "Q4'25",
            "Q1'26",
        ]
    finally:
        workbook.close()


def test_obsolete_valuation_surfaces_and_ghost_cells_are_removed(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        values = [cell.value for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str)]
        for text in (
            "Forward valuation is owned by Investment Case",
            "Forward Valuation Summary",
            "Hidden value flags",
            "Hidden Value Panel",
            "Operating signals",
        ):
            assert text not in values
        assert all(
            cell.value is None and cell.style_id == 0
            for row in sheet.iter_rows(min_row=79, max_row=122, min_col=14, max_col=27)
            for cell in row
        )
        assert all(
            cell.value is None and cell.style_id == 0
            for row in sheet.iter_rows(min_row=169, max_row=188, min_col=1, max_col=13)
            for cell in row
        )
        assert all(
            cell.value is None and cell.style_id == 0
            for row in sheet.iter_rows(min_row=192, max_row=200, min_col=1, max_col=41)
            for cell in row
        )
    finally:
        workbook.close()


def test_formula_retirement_is_exact_and_names_matrix_are_preserved(materialized, plan) -> None:
    output, _ = materialized
    formulas = _formula_map(output)
    assert len(plan.formula_retirement_plan) == 21
    assert sum(row["coordinate"].startswith("Valuation!B") or row["coordinate"].startswith("Valuation!C") or row["coordinate"].startswith("Valuation!D") or row["coordinate"].startswith("Valuation!E") for row in plan.formula_retirement_plan) == 20
    assert not any(key.startswith("Valuation!") for key in formulas)
    assert _defined_names(output) == _defined_names(BASE)
    assert sum(key.split("|", 1)[0].startswith("IC_") for key in _defined_names(output)) == 40
    workbook = load_workbook(output, data_only=False)
    try:
        matrix = workbook["ANF_Investment_Case_Data"]
        assert sum(
            any(matrix.cell(row, column).value is not None for column in range(54, 70))
            for row in range(2, 26)
        ) == 24
    finally:
        workbook.close()


def test_hidden_lineage_relocated_and_reconstructs_all_bindings(materialized, plan) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        records = []
        for row in range(270, 298):
            assert sheet.row_dimensions[row].hidden is True
            value = sheet[f"A{row}"].value
            assert isinstance(value, str)
            records.append(json.loads(value))
        reconstructed = [binding for record in records for binding in record["bindings"]]
        assert reconstructed == list(plan.bindings)
        assert len(reconstructed) == 140
        assert sum(binding["status"] == "available" for binding in reconstructed) == 110
    finally:
        workbook.close()


def test_capital_economics_and_missing_semantics_are_preserved(materialized, plan) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        section_counts = {}
        missing_to_zero = 0
        mismatches = []
        for binding in plan.bindings:
            section_counts.setdefault(binding["section"], [0, 0])
            section_counts[binding["section"]][1] += 1
            if binding["status"] == "available":
                section_counts[binding["section"]][0] += 1
            sheet_name, coordinate = binding["target_cell"].split("!", 1)
            actual = workbook[sheet_name][coordinate].value
            expected = binding["value"]
            if expected is None:
                missing_to_zero += int(actual == 0)
                if actual is not None:
                    mismatches.append(binding["target_cell"])
            elif actual != pytest.approx(float(expected)):
                mismatches.append(binding["target_cell"])
        assert section_counts == {
            "capital_allocation_summary": [12, 12],
            "annual_capital_allocation_history": [14, 20],
            "capital_return_summary": [20, 24],
            "quarterly_capital_return_history": [52, 72],
            "annual_capital_return_history": [12, 12],
        }
        assert missing_to_zero == 0
        assert mismatches == []
    finally:
        workbook.close()


def test_historical_values_do_not_use_manual_input_styling(materialized, plan) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        for binding in plan.bindings:
            coordinate = binding["target_cell"].split("!", 1)[1]
            cell = sheet[coordinate]
            fill_rgb = cell.fill.fgColor.rgb if cell.fill.fill_type else None
            font_rgb = cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None
            assert fill_rgb not in {"00FFF2CC", "FFFFF2CC"}
            assert font_rgb not in {"000070C0", "FF0070C0"}
        assert not any(
            cell.value == "State / definition"
            for row in sheet.iter_rows(min_row=126, max_row=166, min_col=1, max_col=13)
            for cell in row
        )
    finally:
        workbook.close()


def test_unrelated_parts_and_calc_metadata_are_preserved(materialized) -> None:
    output, result = materialized
    with ZipFile(BASE, "r") as before, ZipFile(output, "r") as after:
        valuation_part = _sheet_part_map(before)["Valuation"]
        changed = [
            part
            for part in before.namelist()
            if before.read(part) != after.read(part)
        ]
        assert changed == [valuation_part]
        assert result.changed_ooxml_parts == (valuation_part,)
        assert before.read("xl/workbook.xml") == after.read("xl/workbook.xml")
        assert before.read("xl/styles.xml") == after.read("xl/styles.xml")
        for sheet_name in ("SUMMARY", "BS_Segments", "ANF_Investment_Case", "ANF_Investment_Case_Data", "Hidden_Value_Flags", "Hidden_Value_Audit"):
            part = _sheet_part_map(before)[sheet_name]
            assert before.read(part) == after.read(part)
    assert _defined_names(output) == _defined_names(BASE)


def test_historical_current_grid_is_semantically_unchanged(materialized) -> None:
    output, _ = materialized
    before = load_workbook(BASE, data_only=False)["Valuation"]
    after = load_workbook(output, data_only=False)["Valuation"]
    for row in range(1, 126):
        for column in range(1, 14):
            old = before.cell(row, column)
            new = after.cell(row, column)
            assert (old.value, old.data_type, old.style_id, old.number_format) == (
                new.value,
                new.data_type,
                new.style_id,
                new.number_format,
            )


def test_two_independent_materializations_are_raw_identical(tmp_path: Path, plan) -> None:
    output_a = tmp_path / "a.xlsx"
    output_b = tmp_path / "b.xlsx"
    first = materialize_valuation_capital_product_cleanup(
        plan=plan,
        base_workbook=BASE,
        output_workbook=output_a,
    )
    second = materialize_valuation_capital_product_cleanup(
        plan=plan,
        base_workbook=BASE,
        output_workbook=output_b,
    )
    assert sha256_file(output_a) == sha256_file(output_b)
    assert first.as_dict() == second.as_dict()
