from __future__ import annotations

import hashlib
import json
from pathlib import Path
import re
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
import pytest

from pbi_xbrl.longitudinal_memory.summary_bs_workbook_materialization import (
    _sheet_part_map,
    canonical_ooxml_sha256,
    sha256_file,
)
from pbi_xbrl.longitudinal_memory.valuation_final_investor_polish import (
    ADDITIONAL_NUMERIC_YEAR_HEADER_CELLS,
    ANNUAL_HEADER_ROWS,
    DEBT_HEADER_LABELS,
    DEBT_HEADER_MERGES,
    EXPECTED_BASE_WORKBOOK_SHA256,
    FINAL_VALUATION_DIMENSION,
    INVESTOR_SECTION_SPACER_ROLE,
    MARKET_FORMULAS,
    MARKET_LABELS,
    MARKET_PRICE_OWNER,
    NORMAL_VALUATION_ROW_HEIGHT,
    PERIOD_HEADER_ROWS,
    REMOVED_COMMENT_REFS,
    SPACER_ROWS,
    SUBSECTION_FILL_RGB,
    SUBSECTION_ROWS,
    VALUATION_COLUMN_WIDTH,
    VALUATION_COLUMN_WIDTH_PIXELS,
    build_valuation_final_investor_polish_plan,
    materialize_valuation_final_investor_polish,
)
from pbi_xbrl.longitudinal_memory.valuation_final_layout_cleanup import (
    LINEAGE_SUPPORT_SHEET,
)


DATA_ROOT = Path(r"C:\Users\Jibbe\Aktier\StockModelData")
BASE = (
    DATA_ROOT
    / "audit"
    / "valuation_final_layout_cleanup_2026-08-16"
    / "ANF_valuation_final_layout_cleanup_preview_a.xlsx"
)
NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


@pytest.fixture(scope="session")
def plan():
    assert BASE.is_file(), f"Accepted polish input is unavailable: {BASE}"
    return build_valuation_final_investor_polish_plan(base_workbook=BASE)


@pytest.fixture(scope="session")
def materialized(tmp_path_factory: pytest.TempPathFactory, plan):
    output = tmp_path_factory.mktemp("valuation_final_investor_polish") / "preview.xlsx"
    result = materialize_valuation_final_investor_polish(
        plan=plan, base_workbook=BASE, output_workbook=output
    )
    return output, result


def _formula_map(path: Path) -> dict[str, str]:
    workbook = load_workbook(path, data_only=False)
    try:
        return {
            f"{sheet.title}!{cell.coordinate}": str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        }
    finally:
        workbook.close()


def _defined_names(path: Path) -> dict[str, str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    names = root.find("m:definedNames", NS)
    return {
        f"{node.attrib['name']}|{node.attrib.get('localSheetId', '')}": node.text or ""
        for node in (() if names is None else names)
    }


def _comment_refs(path: Path) -> list[str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/comments/comment2.xml"))
    return [node.attrib["ref"] for node in root.findall("m:commentList/m:comment", NS)]


def _style_rgb(workbook, style_id: int) -> str | None:
    style = workbook._cell_styles[style_id]
    fill = workbook._fills[style.fillId]
    value = fill.fgColor.rgb if fill.fill_type else None
    return None if value is None else value[-6:]


def _lineage_bindings(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=False)
    try:
        support = workbook[LINEAGE_SUPPORT_SHEET]
        records = [json.loads(str(support[f"A{row}"].value)) for row in range(1, 29)]
    finally:
        workbook.close()
    return [binding for record in records for binding in record["bindings"]]


def test_accepted_input_and_plan_are_stable(plan) -> None:
    assert sha256_file(BASE) == EXPECTED_BASE_WORKBOOK_SHA256
    replay = build_valuation_final_investor_polish_plan(base_workbook=BASE)
    assert replay.to_dict() == plan.to_dict()
    assert plan.current_price_owner == MARKET_PRICE_OWNER
    assert plan.market_formula_map == MARKET_FORMULAS
    assert plan.market_disposition == "ACTIVE_CURRENT_MARKET_PRESENTATION_USING_EXISTING_MANUAL_PRICE_INPUT"


def test_final_legacy_comments_are_removed_without_touching_valid_comments(materialized) -> None:
    output, _ = materialized
    before = _comment_refs(BASE)
    after = _comment_refs(output)
    assert set(before) - set(after) == set(REMOVED_COMMENT_REFS)
    assert not set(REMOVED_COMMENT_REFS).intersection(after)
    assert len(after) == len(before) - 2


def test_current_market_price_uses_existing_declarative_input(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        ic = workbook["ANF_Investment_Case"]
        valuation = workbook["Valuation"]
        assert ic["A15"].value == "Current share price"
        assert ic["F15"].value is None
        assert _style_rgb(workbook, ic["F15"].style_id) == "FFF2CC"
        assert ic["G15"].value == '=IF(F15<>"",F15,"")'
        assert valuation["A116"].value == "Market Valuation"
        assert valuation["B117"].value == MARKET_FORMULAS["B117"]
    finally:
        workbook.close()


def test_market_valuation_formula_inventory_is_bounded(materialized) -> None:
    output, result = materialized
    before = _formula_map(BASE)
    after = _formula_map(output)
    delta = {key: value for key, value in after.items() if before.get(key) != value}
    assert delta == {f"Valuation!{key}": value for key, value in MARKET_FORMULAS.items()}
    assert set(before).issubset(after)
    assert result.valuation_formula_count == 7
    assert sum(key.startswith("Valuation!") for key in after) == 7


def test_market_valuation_is_current_only_and_missing_safe(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert [sheet[f"A{row}"].value for row in range(117, 124)] == [
            MARKET_LABELS[row] for row in range(117, 124)
        ]
        assert all(sheet[f"B{row}"].data_type == "f" for row in range(117, 124))
        assert all(
            sheet.cell(row, column).value is None
            for row in range(117, 124)
            for column in range(3, 14)
        )
        assert not any("historical" in str(sheet[f"A{row}"].value).lower() for row in range(116, 124))
    finally:
        workbook.close()


def test_debt_detail_header_uses_final_semantic_column_contract(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert {str(item) for item in sheet.merged_cells.ranges}.issuperset(DEBT_HEADER_MERGES)
        for column, label in DEBT_HEADER_LABELS.items():
            assert sheet[f"{column}126"].value == label
        assert sheet.row_dimensions[126].height == pytest.approx(NORMAL_VALUATION_ROW_HEIGHT)
        visible_anchors = ("A126", "B126", "D126", "E126", "G126", "H126", "J126", "M126")
        assert {sheet[cell].style_id for cell in visible_anchors} == {sheet["A126"].style_id}
        header_style_id = sheet["A126"].style_id
        assert sheet["E126"].style_id == sheet["B126"].style_id
        assert sheet["A126"].alignment.wrap_text is True
        assert sheet["A126"].alignment.vertical == "center"
        assert sheet["N125"].style_id == sheet["M125"].style_id
        assert sheet["O125"].style_id == sheet["M125"].style_id
    finally:
        workbook.close()
    with ZipFile(output, "r") as archive:
        root = ET.fromstring(archive.read("xl/worksheets/sheet2.xml"))
    raw_header = root.find("m:sheetData/m:row[@r='126']", NS)
    assert raw_header is not None
    assert {cell.attrib.get("s") for cell in raw_header.findall("m:c", NS)} == {
        str(header_style_id)
    }
    assert [cell.attrib["r"] for cell in raw_header.findall("m:c", NS)] == [
        f"{column}126" for column in "ABCDEFGHIJKLMNO"
    ]


def test_exact_normal_height_spacer_contract(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert INVESTOR_SECTION_SPACER_ROLE == "investor_section_spacer"
        for row in SPACER_ROWS:
            assert sheet.row_dimensions[row].height == pytest.approx(NORMAL_VALUATION_ROW_HEIGHT)
            assert sheet.row_dimensions[row].hidden is False
            for column in range(1, 14):
                cell = sheet.cell(row, column)
                assert cell.value is None
                assert cell.style_id == 0
                assert cell.comment is None
    finally:
        workbook.close()


def test_annual_headers_are_numeric_integers_not_text(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        expected = {139: [2021, 2022, 2023, 2024, 2025], 168: [2024, 2025]}
        for row in ANNUAL_HEADER_ROWS:
            cells = [sheet.cell(row, column) for column in range(2, len(expected[row]) + 2)]
            assert [cell.value for cell in cells] == expected[row]
            assert all(cell.data_type == "n" for cell in cells)
            assert all(cell.number_format == "0" for cell in cells)
    finally:
        workbook.close()


def test_summary_year_headers_are_numeric_integers_not_text(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert ADDITIONAL_NUMERIC_YEAR_HEADER_CELLS == ("D132", "D147")
        for coordinate in ADDITIONAL_NUMERIC_YEAR_HEADER_CELLS:
            cell = sheet[coordinate]
            assert cell.value == 2025
            assert cell.data_type == "n"
            assert cell.number_format == "0"
            assert cell.alignment.horizontal == "right"
            assert cell.alignment.vertical == "center"
    finally:
        workbook.close()


def test_all_capital_period_headers_are_right_aligned(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        for row, columns in PERIOD_HEADER_ROWS.items():
            assert sheet[f"A{row}"].alignment.horizontal == "left"
            assert sheet[f"A{row}"].alignment.vertical == "center"
            for column in columns:
                assert sheet[f"{column}{row}"].alignment.horizontal == "right"
                assert sheet[f"{column}{row}"].alignment.vertical == "center"
    finally:
        workbook.close()


def test_b_through_m_have_consistent_102_pixel_width(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert VALUATION_COLUMN_WIDTH_PIXELS == 102
        assert all(
            sheet.column_dimensions[column].width == pytest.approx(VALUATION_COLUMN_WIDTH)
            for column in "BCDEFGHIJKLM"
        )
    finally:
        workbook.close()


def test_subsection_shade_has_three_level_contrast(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert _style_rgb(workbook, sheet["A130"].style_id) == "6FA8DC"
        assert {_style_rgb(workbook, sheet[f"A{row}"].style_id) for row in SUBSECTION_ROWS} == {
            SUBSECTION_FILL_RGB
        }
        assert _style_rgb(workbook, sheet["A132"].style_id) == "EAF3FB"
    finally:
        workbook.close()


def test_capital_allocation_flow_and_balance_grouping(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert [sheet[f"A{row}"].value for row in range(133, 138)] == [
            "FCF ($m)",
            "Capex / Reinvestment ($m)",
            "Buybacks ($m)",
            None,
            "Ending net cash / (debt) ($m)",
        ]
    finally:
        workbook.close()


def test_capital_return_summary_final_order_and_two_spacers(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert [sheet[f"A{row}"].value for row in range(148, 158)] == [
            "Buybacks ($m)",
            "Shares repurchased (m)",
            "Avg. repurchase price ($/share)",
            "Dividends ($m)",
            None,
            "Shares issued (m)",
            "Net shares retired / (issued) (m)",
            None,
            "Buybacks / FCF (%)",
            "Authorization remaining ($m)",
        ]
    finally:
        workbook.close()


@pytest.mark.parametrize("start,spacer", [(160, 163), (169, 172)])
def test_quarterly_and_annual_return_history_grouping(materialized, start: int, spacer: int) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert [sheet[f"A{row}"].value for row in range(start, start + 7)] == [
            "Buybacks ($m)",
            "Shares repurchased (m)",
            "Avg. repurchase price ($/share)",
            None,
            "Shares issued (m)",
            "Net shares retired / (issued) (m)",
            "Buybacks / FCF (%)",
        ]
        assert spacer == start + 3
    finally:
        workbook.close()


def test_all_140_capital_bindings_and_110_lineages_survive(materialized) -> None:
    output, result = materialized
    bindings = _lineage_bindings(output)
    assert len(bindings) == 140
    assert sum(item["status"] == "available" for item in bindings) == 110
    assert result.remapped_binding_count == 140
    assert all(str(item["target_cell"]).startswith("Valuation!") for item in bindings)
    assert all(
        item.get("owner") and item.get("source_identity") and item.get("source_ref")
        for item in bindings
        if item["status"] == "available"
    )


def test_capital_binding_values_statuses_and_missing_states_are_unchanged(materialized) -> None:
    output, _ = materialized
    bindings = _lineage_bindings(output)
    workbook = load_workbook(output, data_only=False)
    try:
        mismatches = []
        missing_to_zero = 0
        section_counts: dict[str, list[int]] = {}
        for item in bindings:
            section_counts.setdefault(item["section"], [0, 0])
            section_counts[item["section"]][0] += int(item["status"] == "available")
            section_counts[item["section"]][1] += 1
            sheet_name, coordinate = item["target_cell"].split("!", 1)
            actual = workbook[sheet_name][coordinate].value
            expected = item["value"]
            if expected is None:
                missing_to_zero += int(actual == 0)
                if actual is not None:
                    mismatches.append(item["target_cell"])
            elif actual != pytest.approx(float(expected)):
                mismatches.append(item["target_cell"])
        assert section_counts == {
            "capital_allocation_summary": [12, 12],
            "annual_capital_allocation_history": [14, 20],
            "capital_return_summary": [20, 24],
            "quarterly_capital_return_history": [52, 72],
            "annual_capital_return_history": [12, 12],
        }
        assert mismatches == []
        assert missing_to_zero == 0
    finally:
        workbook.close()


def test_debt_and_liquidity_economics_are_identical(materialized) -> None:
    output, _ = materialized
    before = load_workbook(BASE, data_only=False)
    after = load_workbook(output, data_only=False)
    try:
        for row in (70, 71, 72, 78):
            assert [before["Valuation"].cell(row, column).value for column in range(1, 14)] == [
                after["Valuation"].cell(row, column).value for column in range(1, 14)
            ]
        assert after["Valuation"]["M72"].value == 0
        assert after["Valuation"]["A127"].value == "No funded core debt instruments as of 2026-Q1"
        assert after["Valuation"]["A128"].value == "Leases separate; undrawn ABL remains in liquidity."
    finally:
        before.close()
        after.close()


def test_no_broken_refs_names_or_deleted_surface_references(materialized) -> None:
    output, _ = materialized
    assert _defined_names(output) == _defined_names(BASE)
    with ZipFile(output, "r") as archive:
        assert not any(
            b"#REF!" in archive.read(name)
            for name in archive.namelist()
            if name.endswith((".xml", ".rels", ".vml"))
        )


def test_only_five_authorized_ooxml_parts_change(materialized) -> None:
    output, result = materialized
    with ZipFile(BASE, "r") as before, ZipFile(output, "r") as after:
        support = _sheet_part_map(before)[LINEAGE_SUPPORT_SHEET]
        expected = {
            "xl/comments/comment2.xml",
            "xl/drawings/commentsDrawing2.vml",
            "xl/styles.xml",
            "xl/worksheets/sheet2.xml",
            support,
        }
        assert set(result.changed_ooxml_parts) == expected
        assert set(before.namelist()) == set(after.namelist())
        for name in set(before.namelist()) - expected:
            assert before.read(name) == after.read(name)


def test_calc_metadata_relationships_and_sheet_states_are_preserved(materialized) -> None:
    output, _ = materialized
    with ZipFile(BASE, "r") as before, ZipFile(output, "r") as after:
        assert before.read("xl/workbook.xml") == after.read("xl/workbook.xml")
        assert before.read("xl/_rels/workbook.xml.rels") == after.read("xl/_rels/workbook.xml.rels")
        calc = re.search(rb"<calcPr\b[^>]*/>", after.read("xl/workbook.xml")).group(0)
        assert b'calcMode="auto"' in calc
        assert b'fullCalcOnLoad="1"' in calc
        assert b'forceFullCalc="0"' in calc


def test_final_used_range_is_exact(materialized) -> None:
    output, result = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook["Valuation"].calculate_dimension() == FINAL_VALUATION_DIMENSION
        assert workbook["Valuation"].max_row == 175
    finally:
        workbook.close()
    assert result.valuation_dimension == FINAL_VALUATION_DIMENSION


def test_valuation_worksheet_rows_and_cells_are_strictly_ordered_for_excel(materialized) -> None:
    output, _ = materialized
    with ZipFile(output, "r") as archive:
        root = ET.fromstring(archive.read("xl/worksheets/sheet2.xml"))
    sheet_data = root.find("m:sheetData", NS)
    assert sheet_data is not None

    row_numbers = [int(row.attrib["r"]) for row in sheet_data.findall("m:row", NS)]
    assert row_numbers == sorted(set(row_numbers))
    assert all(row_numbers.count(row) == 1 for row in range(167, 176))

    for row in sheet_data.findall("m:row", NS):
        row_number = int(row.attrib["r"])
        coordinates = [cell.attrib["r"] for cell in row.findall("m:c", NS)]
        assert len(coordinates) == len(set(coordinates))
        assert all(int(re.search(r"[0-9]+$", coordinate).group()) == row_number for coordinate in coordinates)

        def column_index(coordinate: str) -> int:
            letters = re.match(r"[A-Z]+", coordinate).group()
            value = 0
            for letter in letters:
                value = value * 26 + ord(letter) - ord("A") + 1
            return value

        column_indexes = [column_index(coordinate) for coordinate in coordinates]
        assert column_indexes == sorted(column_indexes)


def test_two_independent_materializations_are_deterministic(tmp_path: Path, plan) -> None:
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    first = materialize_valuation_final_investor_polish(plan=plan, base_workbook=BASE, output_workbook=a)
    second = materialize_valuation_final_investor_polish(plan=plan, base_workbook=BASE, output_workbook=b)
    assert sha256_file(a) == sha256_file(b)
    assert canonical_ooxml_sha256(a) == canonical_ooxml_sha256(b)
    assert first.to_dict() == second.to_dict()


def test_artifact_tool_is_not_an_authoring_dependency() -> None:
    source = (
        Path(__file__).resolve().parents[1]
        / "pbi_xbrl"
        / "longitudinal_memory"
        / "valuation_final_investor_polish.py"
    ).read_text(encoding="utf-8")
    assert "artifact inspection/rendering tool" in source
    assert "SpreadsheetFile" not in source
    assert "importXlsx" not in source
    assert "exportXlsx" not in source


def test_no_ticker_branch_or_historical_multiple_engine_was_added() -> None:
    source = (
        Path(__file__).resolve().parents[1]
        / "pbi_xbrl"
        / "longitudinal_memory"
        / "valuation_final_investor_polish.py"
    ).read_text(encoding="utf-8")
    assert 'if ticker == "ANF"' not in source
    assert "historical_multiple" not in source.lower()
    assert hashlib.sha256(source.encode("utf-8")).hexdigest()
