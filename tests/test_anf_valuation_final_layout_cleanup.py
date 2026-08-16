from __future__ import annotations

import json
from pathlib import Path
import re
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
import pytest

from pbi_xbrl.longitudinal_memory.summary_bs_workbook_materialization import (
    _sheet_part_map,
    canonical_ooxml_sha256,
    sha256_file,
)
from pbi_xbrl.longitudinal_memory.valuation_final_layout_cleanup import (
    EXPECTED_BASE_WORKBOOK_SHA256,
    EXPECTED_PRIOR_BINDING_PLAN_DIGEST,
    LINEAGE_SUPPORT_RANGE,
    LINEAGE_SUPPORT_SHEET,
    NORMAL_VALUATION_ROW_HEIGHT,
    OLD_LINEAGE_RANGE,
    RIGHT_SIDE_LEGACY_RANGE,
    build_valuation_final_layout_cleanup_plan,
    materialize_valuation_final_layout_cleanup,
    render_period_label,
)


DATA_ROOT = Path(r"C:\Users\Jibbe\Aktier\StockModelData")
PRIOR_AUDIT = DATA_ROOT / "audit" / "valuation_capital_product_cleanup_2026-08-16"
BASE = PRIOR_AUDIT / "ANF_valuation_capital_product_cleanup_preview_a.xlsx"
PRIOR_PLAN = PRIOR_AUDIT / "work" / "plan.json"
NS = {
    "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def _require_inputs() -> None:
    for path in (BASE, PRIOR_PLAN):
        if not path.is_file():
            pytest.skip(f"Accepted layout input is unavailable: {path}")


@pytest.fixture(scope="session")
def plan():
    _require_inputs()
    return build_valuation_final_layout_cleanup_plan(
        base_workbook=BASE,
        prior_plan_path=PRIOR_PLAN,
    )


@pytest.fixture(scope="session")
def materialized(tmp_path_factory: pytest.TempPathFactory, plan):
    output = tmp_path_factory.mktemp("valuation_final_layout") / "preview.xlsx"
    result = materialize_valuation_final_layout_cleanup(
        plan=plan,
        base_workbook=BASE,
        output_workbook=output,
    )
    return output, result


def _defined_names(path: Path) -> dict[str, str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    names = root.find("m:definedNames", NS)
    return {
        f"{node.attrib['name']}|{node.attrib.get('localSheetId', '')}": node.text or ""
        for node in (() if names is None else names)
    }


def _formula_map(path: Path) -> dict[str, str]:
    workbook = load_workbook(path, data_only=False)
    try:
        return {
            f"{sheet.title}!{cell.coordinate}": str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        }
    finally:
        workbook.close()


def _comment_refs(path: Path) -> list[str]:
    with ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read("xl/comments/comment2.xml"))
    return [node.attrib["ref"] for node in root.findall("m:commentList/m:comment", NS)]


def _style_rgb(workbook, style_id: int) -> str | None:
    style = workbook._cell_styles[style_id]
    fill = workbook._fills[style.fillId]
    value = fill.fgColor.rgb if fill.fill_type else None
    return None if value is None else value[-6:]


def test_accepted_input_and_plan_identity(plan) -> None:
    assert sha256_file(BASE) == EXPECTED_BASE_WORKBOOK_SHA256
    assert plan.base_workbook_sha256 == EXPECTED_BASE_WORKBOOK_SHA256
    assert plan.prior_binding_plan_digest == EXPECTED_PRIOR_BINDING_PLAN_DIGEST
    assert OLD_LINEAGE_RANGE == "Valuation!A270:A297"
    assert LINEAGE_SUPPORT_RANGE == "Capital_Product_Lineage!A1:A28"


def test_plan_replay_is_deterministic(plan) -> None:
    repeat = build_valuation_final_layout_cleanup_plan(
        base_workbook=BASE,
        prior_plan_path=PRIOR_PLAN,
    )
    assert repeat.to_dict() == plan.to_dict()


def test_generic_period_renderer() -> None:
    assert render_period_label("2026-Q1") == "2026-Q1"
    assert render_period_label("2021-FY") == "2021"
    assert render_period_label("TTM through 2026-Q1") == "TTM 2026-Q1"


def test_right_side_legacy_block_is_physically_clean(materialized, plan) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        for row in sheet.iter_rows(min_row=50, max_row=75, min_col=15, max_col=29):
            for cell in row:
                assert cell.value is None
                assert cell.style_id == 0
                assert cell.comment is None
        assert not any(
            not (
                merged.max_col < 15
                or merged.min_col > 29
                or merged.max_row < 50
                or merged.min_row > 75
            )
            for merged in sheet.merged_cells.ranges
        )
    finally:
        workbook.close()
    assert len(plan.right_side_cells) == 390
    assert len(plan.right_side_merges) == 75


def test_row_139_uses_surrounding_normal_height(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert [sheet.cell(139, column).value for column in range(1, 14)] == [None] * 13
        assert sheet.row_dimensions[139].hidden is False
        assert sheet.row_dimensions[139].height == pytest.approx(NORMAL_VALUATION_ROW_HEIGHT)
        assert sheet.row_dimensions[138].height == pytest.approx(NORMAL_VALUATION_ROW_HEIGHT)
        assert sheet.row_dimensions[140].height == pytest.approx(21.0)
    finally:
        workbook.close()


def test_three_level_capital_header_style_contract(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        major = sheet["A126"].style_id
        subsection_ids = {sheet[f"A{row}"].style_id for row in (127, 133, 141, 151, 159)}
        table = sheet["A128"].style_id
        assert major == 90
        assert subsection_ids == {38}
        assert table == 91
        assert len({major, *subsection_ids, table}) == 3
        assert _style_rgb(workbook, major) == "6FA8DC"
        assert _style_rgb(workbook, 38) == "D9E7F3"
        assert _style_rgb(workbook, table) == "EAF3FB"
    finally:
        workbook.close()


def test_all_capital_period_labels_are_normalized(materialized) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert [sheet.cell(128, column).value for column in range(1, 5)] == [
            "Metric", "2026-Q1", "TTM 2026-Q1", "2025"
        ]
        assert [sheet.cell(134, column).value for column in range(1, 7)] == [
            "Metric", "2021", "2022", "2023", "2024", "2025"
        ]
        assert [sheet.cell(142, column).value for column in range(1, 5)] == [
            "Metric", "2026-Q1", "TTM 2026-Q1", "2025"
        ]
        assert [sheet.cell(152, column).value for column in range(1, 14)] == [
            "Metric", "2023-Q2", "2023-Q3", "2023-Q4", "2024-Q1", "2024-Q2",
            "2024-Q3", "2024-Q4", "2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4",
            "2026-Q1",
        ]
        assert [sheet.cell(160, column).value for column in range(1, 4)] == [
            "Metric", "2024", "2025"
        ]
    finally:
        workbook.close()


def test_period_label_plan_preserves_binding_and_value_identity(plan) -> None:
    assert len(plan.period_label_mutations) == 25
    assert all(item.old_label != item.new_label for item in plan.period_label_mutations)
    assert all(re.fullmatch(r"[0-9]{4}(?:-Q[1-4])?", item.new_label) or item.new_label.startswith("TTM ") for item in plan.period_label_mutations)
    assert all(re.fullmatch(r"[0-9a-f]{64}", item.binding_identity_digest) for item in plan.period_label_mutations)
    assert all(re.fullmatch(r"[0-9a-f]{64}", item.value_status_digest) for item in plan.period_label_mutations)


def test_retired_comments_and_red_triangles_are_removed(materialized, plan) -> None:
    output, _ = materialized
    refs = _comment_refs(output)
    assert len(plan.removed_comment_refs) == 109
    assert len(refs) == 40
    assert refs == list(plan.preserved_comment_refs)
    assert not any(192 <= int(re.search(r"[0-9]+", ref).group()) <= 261 for ref in refs)
    assert not any(
        15 <= openpyxl_column(ref) <= 29 and 50 <= openpyxl_row(ref) <= 75
        for ref in refs
    )


def openpyxl_column(coordinate: str) -> int:
    letters = re.match(r"[A-Z]+", coordinate).group()
    result = 0
    for letter in letters:
        result = result * 26 + ord(letter) - 64
    return result


def openpyxl_row(coordinate: str) -> int:
    return int(re.search(r"[0-9]+", coordinate).group())


def test_rows_201_261_are_physically_absent(materialized) -> None:
    output, _ = materialized
    with ZipFile(output, "r") as archive:
        root = ET.fromstring(archive.read("xl/worksheets/sheet2.xml"))
    rows = {int(node.attrib["r"]) for node in root.findall("m:sheetData/m:row", NS)}
    assert not rows.intersection(range(201, 262))
    assert not rows.intersection(range(192, 201))
    assert not rows.intersection(range(270, 298))
    assert not root.findall("m:conditionalFormatting", NS)


def test_lineage_is_byte_preserved_on_dedicated_hidden_sheet(materialized, plan) -> None:
    output, _ = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        valuation = workbook["Valuation"]
        support = workbook[LINEAGE_SUPPORT_SHEET]
        assert support.sheet_state == "hidden"
        assert all(valuation[f"A{row}"].value is None for row in range(270, 298))
        records = [support[f"A{row}"].value for row in range(1, 29)]
        assert records == list(plan.lineage_records)
        assert all(isinstance(record, str) for record in records)
    finally:
        workbook.close()


def test_lineage_reconstructs_140_bindings_and_110_available(materialized) -> None:
    output, _ = materialized
    prior = json.loads(PRIOR_PLAN.read_text(encoding="utf-8"))
    workbook = load_workbook(output, data_only=False)
    try:
        support = workbook[LINEAGE_SUPPORT_SHEET]
        records = [json.loads(support[f"A{row}"].value) for row in range(1, 29)]
    finally:
        workbook.close()
    bindings = [binding for record in records for binding in record["bindings"]]
    assert bindings == prior["bindings"]
    assert len(bindings) == 140
    assert sum(binding["status"] == "available" for binding in bindings) == 110


def test_capital_economics_and_missing_states_are_unchanged(materialized) -> None:
    output, _ = materialized
    prior = json.loads(PRIOR_PLAN.read_text(encoding="utf-8"))
    workbook = load_workbook(output, data_only=False)
    try:
        mismatches = []
        missing_to_zero = 0
        section_counts: dict[str, list[int]] = {}
        for binding in prior["bindings"]:
            section_counts.setdefault(binding["section"], [0, 0])
            section_counts[binding["section"]][1] += 1
            section_counts[binding["section"]][0] += int(binding["status"] == "available")
            sheet_name, coordinate = binding["target_cell"].split("!", 1)
            actual = workbook[sheet_name][coordinate].value
            expected = binding["value"]
            if expected is None:
                missing_to_zero += int(actual == 0)
                if actual is not None:
                    mismatches.append(binding["target_cell"])
            elif actual != pytest.approx(float(expected)):
                mismatches.append(binding["target_cell"])
        assert section_counts == {
            "capital_allocation_summary": [12, 12],
            "annual_capital_allocation_history": [14, 20],
            "capital_return_summary": [20, 24],
            "quarterly_capital_return_history": [52, 72],
            "annual_capital_return_history": [12, 12],
        }
        assert missing_to_zero == 0
        assert mismatches == []
    finally:
        workbook.close()


def test_formula_and_defined_name_references_are_unchanged(materialized) -> None:
    output, _ = materialized
    assert _formula_map(output) == _formula_map(BASE)
    assert _defined_names(output) == _defined_names(BASE)
    assert not any(key.startswith("Valuation!") for key in _formula_map(output))
    with ZipFile(output, "r") as archive:
        assert not any(b"#REF!" in archive.read(name) for name in archive.namelist())


def test_used_range_and_hidden_rows_reflect_simplified_product(materialized) -> None:
    output, result = materialized
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert sheet.calculate_dimension() == "A1:AI166"
        assert sheet.max_row == 166
        assert not [row for row, dimension in sheet.row_dimensions.items() if dimension.hidden]
    finally:
        workbook.close()
    assert result.valuation_dimension == "A1:AI166"


def test_only_authorized_ooxml_parts_change(materialized) -> None:
    output, result = materialized
    expected = {
        "[Content_Types].xml",
        "xl/_rels/workbook.xml.rels",
        "xl/comments/comment2.xml",
        "xl/drawings/commentsDrawing2.vml",
        "xl/workbook.xml",
        "xl/worksheets/sheet2.xml",
        "xl/worksheets/sheet58.xml",
    }
    assert set(result.changed_ooxml_parts) == expected
    with ZipFile(BASE, "r") as before, ZipFile(output, "r") as after:
        for name in set(before.namelist()) - expected:
            assert before.read(name) == after.read(name)
        assert set(after.namelist()) - set(before.namelist()) == {"xl/worksheets/sheet58.xml"}


def test_calc_metadata_is_byte_preserved(materialized) -> None:
    output, _ = materialized
    with ZipFile(BASE, "r") as before, ZipFile(output, "r") as after:
        before_calc = re.search(rb"<calcPr\b[^>]*/>", before.read("xl/workbook.xml")).group(0)
        after_calc = re.search(rb"<calcPr\b[^>]*/>", after.read("xl/workbook.xml")).group(0)
    assert before_calc == after_calc
    assert b'calcMode="auto"' in after_calc
    assert b'fullCalcOnLoad="1"' in after_calc
    assert b'forceFullCalc="0"' in after_calc


def test_two_independent_materializations_are_fully_deterministic(tmp_path: Path, plan) -> None:
    output_a = tmp_path / "a.xlsx"
    output_b = tmp_path / "b.xlsx"
    first = materialize_valuation_final_layout_cleanup(
        plan=plan, base_workbook=BASE, output_workbook=output_a
    )
    second = materialize_valuation_final_layout_cleanup(
        plan=plan, base_workbook=BASE, output_workbook=output_b
    )
    assert sha256_file(output_a) == sha256_file(output_b)
    assert canonical_ooxml_sha256(output_a) == canonical_ooxml_sha256(output_b)
    assert first.to_dict() == second.to_dict()


def test_no_ticker_specific_branch_was_introduced() -> None:
    source = (
        Path(__file__).resolve().parents[1]
        / "pbi_xbrl"
        / "longitudinal_memory"
        / "valuation_final_layout_cleanup.py"
    ).read_text(encoding="utf-8")
    assert 'if ticker == "ANF"' not in source
    assert "if ticker == 'ANF'" not in source


def test_plan_preflight_has_no_surviving_retired_reference(plan) -> None:
    assert plan.reference_preflight == {
        "deleted_or_cleaned_row_reference_count": 0,
        "deleted_or_cleaned_row_references": [],
        "ref_error_part_count": 0,
        "ref_error_parts": [],
    }
    assert RIGHT_SIDE_LEGACY_RANGE == "O50:AC75"
