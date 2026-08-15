from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
import pytest

from pbi_xbrl.longitudinal_memory.formula_aware_workbook_materialization import (
    DefinedNameMutation,
    FormulaAwareCellMutation,
    FormulaAwareMaterializationError,
    WorkbookCalculationMetadataPolicy,
    WorksheetMergeMutation,
    WorksheetRowMutation,
    materialize_formula_aware_mutations,
)


def _base_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Valuation"
    worksheet["A1"] = 1
    worksheet["B1"] = "legacy"
    worksheet["C1"] = 3
    worksheet["A2"] = "title"
    worksheet.merge_cells("A2:C2")
    worksheet["A5"] = "row state"
    workbook.defined_names.add(DefinedName("LegacyName", attr_text="'Valuation'!$A$1"))
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)
    workbook.close()


def test_noop_is_raw_byte_identical(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    output = tmp_path / "noop.xlsx"
    _base_workbook(base)
    result = materialize_formula_aware_mutations(
        base_workbook=base,
        output_workbook=output,
        cell_mutations=(),
    )
    assert base.read_bytes() == output.read_bytes()
    assert result.changed_ooxml_parts == ()
    assert result.cell_mutation_count == 0


def test_formula_value_name_merge_and_row_mutations_are_bounded(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    output = tmp_path / "candidate.xlsx"
    _base_workbook(base)
    result = materialize_formula_aware_mutations(
        base_workbook=base,
        output_workbook=output,
        cell_mutations=(
            FormulaAwareCellMutation("Valuation", "A1", "SET_VALUE", "2", "number"),
            FormulaAwareCellMutation("Valuation", "B1", "SET_FORMULA", "A1*2"),
            FormulaAwareCellMutation("Valuation", "C1", "CLEAR_CONTENTS"),
            FormulaAwareCellMutation("Valuation", "D1", "SET_FORMULA", "A1+1"),
        ),
        defined_name_mutations=(
            DefinedNameMutation("LegacyName", "DELETE"),
            DefinedNameMutation("CanonicalName", "UPSERT", "'Valuation'!$D$1"),
        ),
        merge_mutations=(
            WorksheetMergeMutation("Valuation", "A2:C2", "DELETE"),
            WorksheetMergeMutation("Valuation", "A2:B2", "ADD"),
        ),
        row_mutations=(WorksheetRowMutation("Valuation", 5, True),),
    )
    workbook = load_workbook(output, data_only=False)
    try:
        worksheet = workbook["Valuation"]
        assert worksheet["A1"].value == 2
        assert worksheet["B1"].value == "=A1*2"
        assert worksheet["C1"].value is None
        assert worksheet["D1"].value == "=A1+1"
        assert "A2:B2" in {str(item) for item in worksheet.merged_cells.ranges}
        assert "A2:C2" not in {str(item) for item in worksheet.merged_cells.ranges}
        assert worksheet.row_dimensions[5].hidden is True
        assert workbook.defined_names.get("LegacyName") is None
        assert workbook.defined_names["CanonicalName"].attr_text == "'Valuation'!$D$1"
    finally:
        workbook.close()
    assert result.cell_mutation_count == 4
    assert result.defined_name_delete_count == 1
    assert result.defined_name_upsert_count == 1
    assert result.merge_add_count == 1
    assert result.merge_delete_count == 1
    assert result.row_mutation_count == 1
    with ZipFile(output, "r") as archive:
        worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
    assert b"<f>A1*2</f>" in worksheet_xml
    assert b"<f>A1+1</f>" in worksheet_xml
    assert b"<f>A1*2</f><v>" not in worksheet_xml


def test_formula_external_reference_and_overwrite_fail_closed(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    output = tmp_path / "candidate.xlsx"
    _base_workbook(base)
    with pytest.raises(FormulaAwareMaterializationError, match="External-reference"):
        materialize_formula_aware_mutations(
            base_workbook=base,
            output_workbook=output,
            cell_mutations=(
                FormulaAwareCellMutation("Valuation", "B1", "SET_FORMULA", "[other.xlsx]Sheet1!A1"),
            ),
        )
    output.write_bytes(b"already exists")
    with pytest.raises(FormulaAwareMaterializationError, match="Refusing to overwrite"):
        materialize_formula_aware_mutations(
            base_workbook=base,
            output_workbook=output,
            cell_mutations=(),
        )


def test_calculation_metadata_policy_changes_only_force_full_calc(tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    output = tmp_path / "candidate.xlsx"
    _base_workbook(base)
    policy = WorkbookCalculationMetadataPolicy(
        policy_id="valuation-native-safe-calculation-metadata@1",
    )
    result = materialize_formula_aware_mutations(
        base_workbook=base,
        output_workbook=output,
        cell_mutations=(),
        calculation_metadata_policy=policy,
    )

    with ZipFile(base, "r") as before, ZipFile(output, "r") as after:
        assert before.namelist() == after.namelist()
        changed = [
            name for name in before.namelist() if before.read(name) != after.read(name)
        ]
        assert changed == ["xl/workbook.xml"]
        before_xml = before.read("xl/workbook.xml")
        after_xml = after.read("xl/workbook.xml")
        assert before_xml.replace(b'forceFullCalc="1"', b'forceFullCalc="0"') == after_xml

    assert result.changed_ooxml_parts == ("xl/workbook.xml",)
    assert result.calculation_metadata_change_count == 1
    assert result.calculation_metadata_policy_id == policy.policy_id
    assert result.calculation_metadata_before is not None
    assert result.calculation_metadata_after is not None
    assert result.calculation_metadata_before["calcMode"] == "auto"
    assert result.calculation_metadata_before["fullCalcOnLoad"] == "1"
    assert result.calculation_metadata_before["forceFullCalc"] == "1"
    assert result.calculation_metadata_after == {
        **result.calculation_metadata_before,
        "forceFullCalc": "0",
    }


def test_calculation_metadata_policy_fails_closed_if_precondition_changed(
    tmp_path: Path,
) -> None:
    base = tmp_path / "base.xlsx"
    output = tmp_path / "candidate.xlsx"
    _base_workbook(base)
    workbook = load_workbook(base)
    workbook.calculation.forceFullCalc = False
    workbook.save(base)
    workbook.close()

    with pytest.raises(FormulaAwareMaterializationError, match="precondition mismatch"):
        materialize_formula_aware_mutations(
            base_workbook=base,
            output_workbook=output,
            cell_mutations=(),
            calculation_metadata_policy=WorkbookCalculationMetadataPolicy(
                policy_id="valuation-native-safe-calculation-metadata@1",
            ),
        )
