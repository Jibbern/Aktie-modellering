from __future__ import annotations

from copy import deepcopy
import json
from pathlib import Path
import re
from zipfile import ZipFile

from openpyxl import load_workbook
import pytest

from pbi_xbrl.longitudinal_memory.summary_bs_workbook_materialization import (
    _cell_elements,
    _sheet_part_map,
    sha256_file,
)
from pbi_xbrl.longitudinal_memory.valuation_final_investor_polish import (
    COMMENTS_PART,
    COMMENTS_VML_PART,
    MARKET_FORMULAS,
    VALUATION_PART,
    _formula_text,
)
from pbi_xbrl.longitudinal_memory.valuation_final_layout_cleanup import (
    LINEAGE_SUPPORT_SHEET,
    _COMMENT_RE,
    _attributes,
    _comment_reference,
    _inline_text,
)
from pbi_xbrl.longitudinal_memory.valuation_guidance_net_share_polish import (
    ANNUAL_PERCENTAGE_ROW,
    ANNUAL_SPACER_ROW,
    GUIDANCE_COMMENT_MOVES,
    NET_SHARE_PERCENTAGE_DEFINITION,
    NET_SHARE_PERCENTAGE_FORMAT,
    NET_SHARE_PERCENTAGE_LABEL,
    NET_SHARE_PERCENTAGE_METRIC_ID,
    OPERATING_DRIVER_COMMENT_REMOVALS,
    SUMMARY_PERCENTAGE_ROW,
    SUMMARY_SPACER_ROW,
    ValuationGuidanceNetSharePolishError,
    _guidance_fit_inventory,
    _read_support_records,
    build_valuation_guidance_net_share_polish_plan,
    derive_net_share_percentage_records,
    materialize_valuation_guidance_net_share_polish,
)


ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = Path(r"C:\Users\Jibbe\Aktier\StockModelData")
BASE = (
    DATA_ROOT
    / "audit"
    / "valuation_header_polish_fix_2026-08-16_final"
    / "ANF_valuation_final_investor_polish_preview_a.xlsx"
)
PACKAGE_PATH = (
    DATA_ROOT
    / "outputs"
    / "stress_tests"
    / "ANF_new_ticker_engine"
    / "ANF_normalized_data_package.json"
)
BS_PRODUCT_PATH = ROOT / "tests" / "fixtures" / "summary_bs" / "anf_bs_segment_product.v1.json"
BS_SHADOW_PATH = ROOT / "tests" / "fixtures" / "summary_bs" / "anf_bs_segment_shadow.v1.json"
EXPECTED_CHANGED_PARTS = {
    VALUATION_PART,
    COMMENTS_PART,
    COMMENTS_VML_PART,
    "xl/worksheets/sheet58.xml",
}


def _load(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def source_inputs() -> tuple[dict, dict, dict]:
    return _load(PACKAGE_PATH), _load(BS_PRODUCT_PATH), _load(BS_SHADOW_PATH)


@pytest.fixture(scope="module")
def support_records() -> tuple[dict, ...]:
    with ZipFile(BASE, "r") as archive:
        support_part = _sheet_part_map(archive)[LINEAGE_SUPPORT_SHEET]
        return _read_support_records(archive.read(support_part))


@pytest.fixture(scope="module")
def plan(source_inputs: tuple[dict, dict, dict]):
    package, product, shadow = source_inputs
    return build_valuation_guidance_net_share_polish_plan(
        base_workbook=BASE,
        source_package=package,
        source_package_path=PACKAGE_PATH,
        balance_sheet_product=product,
        balance_sheet_product_path=BS_PRODUCT_PATH,
        balance_sheet_shadow=shadow,
        balance_sheet_shadow_path=BS_SHADOW_PATH,
    )


@pytest.fixture(scope="module")
def candidate(plan, tmp_path_factory: pytest.TempPathFactory) -> Path:
    output = tmp_path_factory.mktemp("valuation-guidance-polish") / "candidate.xlsx"
    materialize_valuation_guidance_net_share_polish(
        plan=plan,
        base_workbook=BASE,
        output_workbook=output,
    )
    return output


def _percentage_bindings(records: tuple[dict, dict]) -> list[dict]:
    return [binding for record in records for binding in record["bindings"]]


def _remove_bs_period(product: dict, period_id: str) -> dict:
    changed = deepcopy(product)
    changed["fields"] = [
        field
        for field in changed["fields"]
        if not (
            field.get("metric_key") == "shares_outstanding"
            and field.get("period_id") == period_id
        )
    ]
    return changed


def _mutate_capital_numerator(package: dict, support: tuple[dict, ...], period: str, value: float):
    changed_package = deepcopy(package)
    changed_support = deepcopy(support)
    record_id = None
    for record in changed_package["capital_returns"]["records"]:
        if record.get("metric_id") == "net_share_reduction" and record.get("fiscal_period") == period:
            record["value"] = value
            record_id = record["record_id"]
            break
    assert record_id
    for record in changed_support:
        for binding in record["bindings"]:
            if binding.get("source_identity") == record_id:
                binding["value"] = value
    return changed_package, changed_support


def test_plan_preserves_baseline_and_expands_only_five_bindings(plan) -> None:
    assert plan.old_binding_count == 140
    assert plan.old_available_binding_count == 110
    assert plan.new_binding_count == 145
    assert plan.new_available_binding_count == 114
    assert plan.new_unavailable_binding_count == 31
    assert plan.added_metric_instance_count == 5
    assert plan.prior_binding_plan_digest != plan.binding_plan_digest
    assert plan.guidance_fit_inventory == {
        "guidance_capacity_width_units": 64.0,
        "guidance_max_text_length": 31,
        "trend_capacity_width_units": 118.0,
        "trend_max_text_length": 35,
    }


def test_net_share_percentage_definition_and_exact_values(plan) -> None:
    bindings = _percentage_bindings(plan.net_share_percentage_records)
    by_period = {binding["period"]: binding for binding in bindings}
    assert by_period["2026-Q1"]["value"] == pytest.approx(0.574 / 45.005)
    assert by_period["2026-Q1"]["denominator_period"] == "2025-Q4"
    assert by_period["TTM through 2026-Q1"]["value"] == pytest.approx(3.212 / 47.643)
    assert by_period["TTM through 2026-Q1"]["denominator_period"] == "2025-Q1"
    assert by_period["2025-FY"]["value"] == pytest.approx(4.730 / 49.735)
    assert by_period["2025-FY"]["denominator_period"] == "2024-Q4"
    assert by_period["2024-FY"]["value"] is None
    assert by_period["2024-FY"]["status"] == "unavailable"
    assert by_period["2024-FY"]["denominator_period"] == "2023-Q4"
    assert all(binding["definition"] == NET_SHARE_PERCENTAGE_DEFINITION for binding in bindings)
    assert all(binding["number_format"] == NET_SHARE_PERCENTAGE_FORMAT for binding in bindings)


def test_all_available_percentages_have_typed_lineage(plan) -> None:
    for binding in _percentage_bindings(plan.net_share_percentage_records):
        if binding["status"] != "available":
            continue
        assert binding["numerator_field_id"]
        assert binding["denominator_field_id"].startswith("product-field:v1|")
        assert binding["denominator_canonical_fact_id"].startswith("canonical-fact:v1|")
        assert binding["denominator_audit_field_id"].startswith("audit-field:v1|")
        assert binding["period_compatibility"]
        assert binding["derivation_rule"] == "net_share_reduction_m / beginning_period_end_shares_m"
        assert len(binding["component_source_identities"]) == 2


def test_missing_opening_share_fails_closed(
    source_inputs: tuple[dict, dict, dict], support_records: tuple[dict, ...]
) -> None:
    package, product, shadow = source_inputs
    product = _remove_bs_period(product, "period:anf:fy2025-q4@1")
    summary, _ = derive_net_share_percentage_records(
        support_records=support_records,
        package=package,
        balance_sheet_product=product,
        balance_sheet_shadow=shadow,
    )
    q1 = next(binding for binding in summary["bindings"] if binding["period"] == "2026-Q1")
    assert q1["status"] == "unavailable"
    assert q1["value"] is None


def test_wrong_period_opening_share_is_not_substituted(
    source_inputs: tuple[dict, dict, dict], support_records: tuple[dict, ...]
) -> None:
    package, product, shadow = source_inputs
    product = _remove_bs_period(product, "period:anf:fy2025-q1@1")
    summary, _ = derive_net_share_percentage_records(
        support_records=support_records,
        package=package,
        balance_sheet_product=product,
        balance_sheet_shadow=shadow,
    )
    ttm = next(binding for binding in summary["bindings"] if binding["period"].startswith("TTM"))
    assert ttm["status"] == "unavailable"
    assert ttm["denominator_period"] == "2025-Q1"


def test_weighted_average_shares_are_rejected_as_denominator(
    source_inputs: tuple[dict, dict, dict], support_records: tuple[dict, ...]
) -> None:
    package, product, shadow = deepcopy(source_inputs)
    for field in product["fields"]:
        if field.get("metric_key") == "shares_outstanding" and field.get("period_id") == "period:anf:fy2024-q4@1":
            field["semantic_role"] = "weighted_average"
    summary, annual = derive_net_share_percentage_records(
        support_records=support_records,
        package=package,
        balance_sheet_product=product,
        balance_sheet_shadow=shadow,
    )
    assert next(row for row in summary["bindings"] if row["period"] == "2025-FY")["status"] == "unavailable"
    assert next(row for row in annual["bindings"] if row["period"] == "2025-FY")["status"] == "unavailable"


@pytest.mark.parametrize("numerator", [-0.5, 0.0])
def test_negative_issuance_and_zero_change_keep_sign_semantics(
    numerator: float,
    source_inputs: tuple[dict, dict, dict],
    support_records: tuple[dict, ...],
) -> None:
    package, product, shadow = source_inputs
    changed_package, changed_support = _mutate_capital_numerator(
        package, support_records, "2026-Q1", numerator
    )
    summary, _ = derive_net_share_percentage_records(
        support_records=changed_support,
        package=changed_package,
        balance_sheet_product=product,
        balance_sheet_shadow=shadow,
    )
    q1 = next(binding for binding in summary["bindings"] if binding["period"] == "2026-Q1")
    assert q1["value"] == pytest.approx(numerator / 45.005)
    assert (q1["value"] < 0) is (numerator < 0)


def test_competing_opening_share_values_fail_as_ownership_conflict(
    source_inputs: tuple[dict, dict, dict], support_records: tuple[dict, ...]
) -> None:
    package, product, shadow = deepcopy(source_inputs)
    for field in product["fields"]:
        if field.get("metric_key") == "shares_outstanding" and field.get("period_id") == "period:anf:fy2025-q4@1":
            field["value"]["value"] = 45.5
    with pytest.raises(ValuationGuidanceNetSharePolishError, match="ownership conflict"):
        derive_net_share_percentage_records(
            support_records=support_records,
            package=package,
            balance_sheet_product=product,
            balance_sheet_shadow=shadow,
        )


def test_guidance_fit_contract_rejects_clipping_mutation() -> None:
    with ZipFile(BASE, "r") as source:
        mutated = source.read(VALUATION_PART).replace(b"+3-5%", b"X" * 80, 1)
    with pytest.raises(ValuationGuidanceNetSharePolishError, match="clip"):
        _guidance_fit_inventory(mutated)


def test_operating_drivers_visible_surface_and_comments_are_removed(candidate: Path) -> None:
    workbook = load_workbook(candidate, data_only=False)
    try:
        sheet = workbook["Valuation"]
        survivors = [
            cell.coordinate
            for row in sheet.iter_rows(min_row=37, max_row=46, min_col=15, max_col=29)
            for cell in row
            if cell.value is not None or cell.style_id or cell.comment is not None
        ]
        assert survivors == []
        assert not any(
            merged.min_row <= 46 and merged.max_row >= 37 and merged.min_col <= 29 and merged.max_col >= 15
            for merged in sheet.merged_cells.ranges
        )
        assert workbook["Operating_Drivers"].max_row > 1
    finally:
        workbook.close()
    with ZipFile(candidate, "r") as archive:
        refs = {
            _comment_reference(match.group(0))
            for match in _COMMENT_RE.finditer(archive.read(COMMENTS_PART))
        }
    assert not set(OPERATING_DRIVER_COMMENT_REMOVALS) & refs


def test_guidance_is_preserved_and_compressed(candidate: Path) -> None:
    before = load_workbook(BASE, data_only=False)
    after = load_workbook(candidate, data_only=False)
    try:
        left = before["Valuation"]
        right = after["Valuation"]
        for row in list(range(8, 26)) + list(range(28, 36)):
            assert left[f"O{row}"].value == right[f"O{row}"].value
            assert left[f"Q{row}"].value == right[f"Q{row}"].value
            assert left[f"R{row}"].value == right[f"R{row}"].value
            assert left[f"S{row}"].value == right[f"S{row}"].value
            assert left[f"AA{row}"].value == right[f"W{row}"].value
            assert right[f"AA{row}"].value is None
        merges = {str(item) for item in right.merged_cells.ranges}
        assert "S9:V9" in merges
        assert "W9:AC9" in merges
        assert "S9:Z9" not in merges
        assert "AA9:AC9" not in merges
    finally:
        before.close()
        after.close()
    with ZipFile(candidate, "r") as archive:
        refs = {
            _comment_reference(match.group(0))
            for match in _COMMENT_RE.finditer(archive.read(COMMENTS_PART))
        }
    assert set(GUIDANCE_COMMENT_MOVES.values()) <= refs
    assert not set(GUIDANCE_COMMENT_MOVES) & refs


def test_capital_return_summary_and_annual_layout(candidate: Path) -> None:
    workbook = load_workbook(candidate, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert [sheet[f"A{row}"].value for row in range(148, 159)] == [
            "Buybacks ($m)",
            "Shares repurchased (m)",
            "Avg. repurchase price ($/share)",
            "Dividends ($m)",
            None,
            "Shares issued (m)",
            "Net shares retired / (issued) (m)",
            NET_SHARE_PERCENTAGE_LABEL,
            None,
            "Buybacks / FCF (%)",
            "Authorization remaining ($m)",
        ]
        assert [sheet[f"A{row}"].value for row in range(170, 179)] == [
            "Buybacks ($m)",
            "Shares repurchased (m)",
            "Avg. repurchase price ($/share)",
            None,
            "Shares issued (m)",
            "Net shares retired / (issued) (m)",
            NET_SHARE_PERCENTAGE_LABEL,
            None,
            "Buybacks / FCF (%)",
        ]
        assert sheet.row_dimensions[SUMMARY_SPACER_ROW].height == pytest.approx(19.5)
        assert sheet.row_dimensions[ANNUAL_SPACER_ROW].height == pytest.approx(19.5)
        assert sheet[f"B{SUMMARY_PERCENTAGE_ROW}"].number_format == NET_SHARE_PERCENTAGE_FORMAT
        assert sheet[f"C{ANNUAL_PERCENTAGE_ROW}"].number_format == NET_SHARE_PERCENTAGE_FORMAT
        assert sheet[f"B{ANNUAL_PERCENTAGE_ROW}"].value is None
    finally:
        workbook.close()


def test_quarterly_history_keeps_six_metrics_and_no_percentage(candidate: Path) -> None:
    workbook = load_workbook(candidate, data_only=False)
    try:
        sheet = workbook["Valuation"]
        assert [sheet[f"A{row}"].value for row in range(161, 168)] == [
            "Buybacks ($m)",
            "Shares repurchased (m)",
            "Avg. repurchase price ($/share)",
            None,
            "Shares issued (m)",
            "Net shares retired / (issued) (m)",
            "Buybacks / FCF (%)",
        ]
        assert all(sheet[f"A{row}"].value != NET_SHARE_PERCENTAGE_LABEL for row in range(159, 168))
    finally:
        workbook.close()


def test_capital_allocation_is_unchanged(candidate: Path) -> None:
    before = load_workbook(BASE, data_only=False)
    after = load_workbook(candidate, data_only=False)
    try:
        left = before["Valuation"]
        right = after["Valuation"]
        for row in range(130, 144):
            for column in range(1, 14):
                left_cell = left.cell(row, column)
                right_cell = right.cell(row, column)
                assert (left_cell.value, left_cell.style_id, left_cell.number_format) == (
                    right_cell.value,
                    right_cell.style_id,
                    right_cell.number_format,
                )
    finally:
        before.close()
        after.close()


def test_binding_readback_and_hidden_lineage_are_complete(candidate: Path, plan) -> None:
    workbook = load_workbook(candidate, data_only=False)
    try:
        valuation = workbook["Valuation"]
        support = workbook[LINEAGE_SUPPORT_SHEET]
        records = [json.loads(str(support[f"A{row}"].value)) for row in range(1, 31)]
        bindings = [binding for record in records for binding in record["bindings"]]
        assert len(bindings) == 145
        assert sum(row["status"] == "available" for row in bindings) == 114
        assert sum(row["status"] != "available" for row in bindings) == 31
        assert support.sheet_state == "hidden"
        for binding in bindings:
            target = binding["target_cell"].split("!", 1)[1]
            expected = binding["value"] if binding["status"] == "available" else None
            assert valuation[target].value == expected
            if binding["status"] == "available":
                assert binding["source_identity"]
        assert not any(
            isinstance(cell.value, str) and "source_identity" in cell.value
            for row in valuation.iter_rows(min_row=1, max_row=178, min_col=1, max_col=35)
            for cell in row
        )
    finally:
        workbook.close()


def test_formula_references_and_calc_metadata_are_unchanged(candidate: Path) -> None:
    with ZipFile(BASE, "r") as left, ZipFile(candidate, "r") as right:
        assert left.read("xl/workbook.xml") == right.read("xl/workbook.xml")
        assert left.read("xl/styles.xml") == right.read("xl/styles.xml")
        assert left.read("xl/_rels/workbook.xml.rels") == right.read("xl/_rels/workbook.xml.rels")
        candidate_formula_texts = [
            _formula_text(value[2])
            for value in _cell_elements(right.read(VALUATION_PART)).values()
            if _formula_text(value[2]) is not None
        ]
        base_formula_texts = [
            _formula_text(value[2])
            for value in _cell_elements(left.read(VALUATION_PART)).values()
            if _formula_text(value[2]) is not None
        ]
        assert candidate_formula_texts == base_formula_texts
        assert len(candidate_formula_texts) == len(MARKET_FORMULAS) == 7
        assert b"#REF!" not in right.read(VALUATION_PART)


def test_only_authorized_ooxml_parts_change(candidate: Path) -> None:
    with ZipFile(BASE, "r") as left, ZipFile(candidate, "r") as right:
        assert left.namelist() == right.namelist()
        changed = {
            name for name in left.namelist() if left.read(name) != right.read(name)
        }
    assert changed == EXPECTED_CHANGED_PARTS


def test_no_trend_heatmap_or_anf_only_economic_branch(candidate: Path) -> None:
    source = (
        ROOT
        / "pbi_xbrl"
        / "longitudinal_memory"
        / "valuation_guidance_net_share_polish.py"
    ).read_text(encoding="utf-8")
    assert "if ticker" not in source.lower()
    assert "conditionalFormatting" not in source
    with ZipFile(BASE, "r") as left, ZipFile(candidate, "r") as right:
        before = left.read(VALUATION_PART)
        after = right.read(VALUATION_PART)
    assert before.count(b"<conditionalFormatting") == after.count(b"<conditionalFormatting")


def test_deterministic_replay(plan, tmp_path: Path) -> None:
    output_a = tmp_path / "a.xlsx"
    output_b = tmp_path / "b.xlsx"
    result_a = materialize_valuation_guidance_net_share_polish(
        plan=plan, base_workbook=BASE, output_workbook=output_a
    )
    result_b = materialize_valuation_guidance_net_share_polish(
        plan=plan, base_workbook=BASE, output_workbook=output_b
    )
    assert sha256_file(output_a) == sha256_file(output_b)
    assert result_a.to_dict() == result_b.to_dict()
