from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
import pytest

from pbi_xbrl.longitudinal_memory.formula_aware_workbook_materialization import (
    materialize_formula_aware_mutations,
)
from pbi_xbrl.longitudinal_memory.valuation_source_native_projection import (
    BASE_SUMMARY_BS_GOLDEN_SHA256,
    CALCULATION_METADATA_POLICY_ID,
    VALUATION_CALCULATION_METADATA_POLICY,
    build_valuation_projection_plan,
)
from pbi_xbrl.new_ticker_investment_case_formula_surface import (
    canonical_investment_case_defined_names,
)


DATA_ROOT = Path(r"C:\Users\Jibbe\Aktier\StockModelData")
BASE = DATA_ROOT / "audit" / "summary_bs_golden_acceptance_2026-08-14" / "golden" / "ANF_summary_bs_source_native_golden_v1.xlsx"
DONOR = DATA_ROOT / "outputs" / "stress_tests" / "ANF_new_ticker_engine" / "ANF_shadow_model_passB2_four_p1_correction_preview.xlsx"
AUDIT = DATA_ROOT / "audit" / "valuation_exhaustive_reconciliation_2026-08-15"


def _require_inputs() -> None:
    missing = [str(path) for path in (BASE, DONOR, AUDIT) if not path.exists()]
    if missing:
        pytest.skip(f"Local accepted Valuation inputs are unavailable: {missing!r}")


def test_projection_plan_is_exactly_bounded() -> None:
    _require_inputs()
    plan = build_valuation_projection_plan(
        base_workbook=BASE,
        donor_workbook=DONOR,
        exhaustive_audit_dir=AUDIT,
    )
    assert plan.base_workbook_sha256 == BASE_SUMMARY_BS_GOLDEN_SHA256
    assert len(plan.liquidity_issue_cells) == 47
    assert len(plan.securities_net_cash_issue_cells) == 5
    assert len(plan.interest_coverage_cells) == 13
    assert len(plan.old_formula_retirement_cells) == 74
    assert len(plan.compact_link_cells) == 20
    assert len(canonical_investment_case_defined_names()) == 40
    assert plan.ic_dependency_closure.seed_count == 412
    assert plan.ic_dependency_closure.cell_count == 1346
    assert plan.ic_dependency_closure.cells_by_sheet == {
        "ANF_Investment_Case": 146,
        "ANF_Investment_Case_Data": 1200,
    }
    assert len(plan.merge_mutations) == 8
    assert len(plan.row_mutations) == 61


def test_full_projection_closes_bounded_failures(tmp_path: Path) -> None:
    _require_inputs()
    output = tmp_path / "valuation_preview.xlsx"
    plan = build_valuation_projection_plan(
        base_workbook=BASE,
        donor_workbook=DONOR,
        exhaustive_audit_dir=AUDIT,
    )
    result = materialize_formula_aware_mutations(
        base_workbook=BASE,
        output_workbook=output,
        cell_mutations=plan.cell_mutations,
        defined_name_mutations=plan.defined_name_mutations,
        merge_mutations=plan.merge_mutations,
        row_mutations=plan.row_mutations,
        calculation_metadata_policy=VALUATION_CALCULATION_METADATA_POLICY,
        expected_base_sha256=plan.base_workbook_sha256,
    )
    workbook = load_workbook(output, data_only=False)
    try:
        valuation = workbook["Valuation"]
        formulas = {
            cell.coordinate: cell.value
            for row in valuation.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        }
        assert len(formulas) == 21
        assert formulas["AI139"] == '=IFERROR(MATCH(1,\'Hidden_Value_Flags\'!$L$2:$L$100,0)+1,"")'
        assert formulas["B194"] == "=IC_Current_GAAP_EPS"
        assert formulas["C194"] == "=IC_Bear_GAAP_EPS"
        assert formulas["D194"] == "=IC_Base_GAAP_EPS"
        assert formulas["E194"] == "=IC_Bull_GAAP_EPS"
        assert valuation["M94"].value == pytest.approx(0.469)
        assert valuation["M95"].value == pytest.approx(449.531)
        assert valuation["M96"].value == pytest.approx(1043.611)
        assert valuation["M71"].value == pytest.approx(25.144)
        assert valuation["M78"].value == pytest.approx(619.224)
        assert all(valuation[f"{column}88"].value is None for column in "BCDEFGHIJKLM")
        assert valuation["B149"].value == "Interest coverage unavailable under the accepted definition."
        assert valuation["E243"].value is None
        assert valuation["E244"].value is None
        assert valuation["A116"].value == "Market-linked — unavailable (current price not populated)"
        assert valuation["A124"].value == "No funded core debt instruments as of 2026-Q1"
        assert all(valuation.row_dimensions[row].hidden for row in range(201, 262))
        assert "A192:F192" in {str(item) for item in valuation.merged_cells.ranges}
        for name, (sheet, coordinate) in canonical_investment_case_defined_names().items():
            reference = workbook.defined_names[name].attr_text
            resolved_sheet = sheet.replace("{ticker}", "ANF")
            column = "".join(character for character in coordinate if character.isalpha())
            row = "".join(character for character in coordinate if character.isdigit())
            assert reference == f"'{resolved_sheet}'!${column}${row}"
    finally:
        workbook.close()
    assert result.write_type_counts["formula"] == 370
    assert result.calculation_metadata_change_count == 1
    assert result.calculation_metadata_policy_id == CALCULATION_METADATA_POLICY_ID
    with ZipFile(output, "r") as archive:
        workbook_xml = archive.read("xl/workbook.xml")
    assert b'calcMode="auto"' in workbook_xml
    assert b'fullCalcOnLoad="1"' in workbook_xml
    assert b'forceFullCalc="0"' in workbook_xml
    assert b'forceFullCalc="1"' not in workbook_xml
    assert result.changed_ooxml_parts == (
        "xl/styles.xml",
        "xl/workbook.xml",
        "xl/worksheets/sheet2.xml",
        "xl/worksheets/sheet39.xml",
        "xl/worksheets/sheet5.xml",
    )
