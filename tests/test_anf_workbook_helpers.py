import datetime as dt

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from pbi_xbrl.excel_writer_context import (
    ANF_SEGMENT_BRAND_EXPLANATION,
    _anf_add_total_company_quarter_revenue_from_history,
    _anf_annual_segment_data_from_slides_segments,
    _anf_buyback_execution_is_year_or_ttm,
    _anf_build_promise_progress_sections,
    _anf_build_investment_case_data,
    _anf_clear_valuation_side_panels,
    _anf_clean_visible_ui_text,
    _anf_clean_visible_operating_driver_records,
    _anf_compact_driver_group,
    _anf_compact_driver_label,
    _anf_financial_schedule_support_doc_for_quarter,
    _anf_guidance_visible_period_label,
    _anf_investment_case_sheet_order,
    _anf_normalize_qa_status_rows,
    _anf_normalize_ytd_buyback_cash_map_for_valuation,
    _anf_format_year_ttm_buyback_summary,
    _anf_polish_quarter_note_visible_fields,
    _anf_prior_year_quarter,
    _anf_recent_operating_commentary_rows,
    _anf_round_visible_driver_value,
    _anf_value_delta_map_for_fiscal_periods,
    _anf_valuation_guidance_rows,
    _anf_visible_guidance_normalized_frame,
    _anf_visible_quarter_label,
    _anf_visible_quarter_note_summaries,
    _anf_yoy_map_for_fiscal_periods,
    _investment_case_sheet_order,
    _net_debt_yoy_flag_label_and_status_for_position,
    _sector_build_investment_case_data,
    _shared_readable_source_type_label,
    _sector_operating_driver_intro_tables,
    _source_backed_debt_tranches_from_slides,
    _standardize_quarter_notes_ui_categories,
    _rewrite_shared_promise_progress_ui_from_blocks,
    _write_anf_valuation_side_panel,
    _write_anf_investment_case_sheet,
    _write_anf_investment_case_data_sheet,
    _write_sector_investment_case_sheet,
    _write_sector_investment_case_data_sheet,
    _annual_segment_latest_year_for_qa,
    _apply_shared_ui_conventions_to_workbook,
    _filter_anf_quarterly_segment_actual_rows,
    _slides_guidance_has_explicit_metric,
)
from pbi_xbrl.quarter_notes import validate_quarter_notes


def test_anf_annual_segment_data_from_slides_segments_includes_fy2025_regions() -> None:
    slides_segments = pd.DataFrame(
        [
            {"quarter": "2026-01-31", "segment": "Americas", "metric": "revenue", "value": 4_290_395_000.0, "period_type": "annual", "doc": "ANF_2025_annual_report.pdf"},
            {"quarter": "2026-01-31", "segment": "EMEA", "metric": "revenue", "value": 818_140_000.0, "period_type": "annual", "doc": "ANF_2025_annual_report.pdf"},
            {"quarter": "2026-01-31", "segment": "APAC", "metric": "revenue", "value": 157_757_000.0, "period_type": "annual", "doc": "ANF_2025_annual_report.pdf"},
        ]
    )

    out = _anf_annual_segment_data_from_slides_segments(slides_segments)

    assert out["metrics"]["Revenues"]["Americas"][2025] == pytest.approx(4_290_395_000.0)
    assert out["metrics"]["Revenues"]["EMEA"][2025] == pytest.approx(818_140_000.0)
    assert out["metrics"]["Revenues"]["APAC"][2025] == pytest.approx(157_757_000.0)


def test_anf_quarterly_segment_filter_excludes_annual_and_tiny_revenue_rows() -> None:
    slides_segments = pd.DataFrame(
        [
            {"quarter": "2026-01-31", "segment": "Americas", "metric": "revenue", "value": 4_290_395_000.0, "period_type": "annual"},
            {"quarter": "2025-11-01", "segment": "APAC", "metric": "revenue", "value": 26.0, "period_type": "quarter"},
            {"quarter": "2025-11-01", "segment": "EMEA", "metric": "revenue", "value": 194_510_000.0, "period_type": "quarter"},
        ]
    )

    out = _filter_anf_quarterly_segment_actual_rows(slides_segments)

    assert list(out["segment"]) == ["EMEA"]
    assert out.iloc[0]["value"] == pytest.approx(194_510_000.0)


def test_anf_quarterly_segment_filter_uses_history_revenue_anchor_to_reject_annual_leaks() -> None:
    q4_2024 = dt.date(2025, 2, 1)
    q4_2025 = dt.date(2026, 1, 31)
    slides_segments = pd.DataFrame(
        [
            # These are FY2024 annual brand/total values incorrectly tagged as quarter rows.
            {"quarter": q4_2024, "segment": "Total Company", "metric": "revenue", "value": 4_948_587_000.0, "period_type": "quarter"},
            {"quarter": q4_2024, "segment": "Abercrombie", "metric": "revenue", "value": 2_556_434_000.0, "period_type": "quarter"},
            {"quarter": q4_2024, "segment": "Hollister", "metric": "revenue", "value": 2_392_153_000.0, "period_type": "quarter"},
            # These are true Q4 geography values and should stay.
            {"quarter": q4_2024, "segment": "Americas", "metric": "revenue", "value": 1_319_720_000.0, "period_type": "quarter"},
            {"quarter": q4_2024, "segment": "EMEA", "metric": "revenue", "value": 224_467_000.0, "period_type": "quarter"},
            {"quarter": q4_2024, "segment": "APAC", "metric": "revenue", "value": 40_730_000.0, "period_type": "quarter"},
            # Latest Q4 FY2025 brand values tie to quarterly revenue and should stay.
            {"quarter": q4_2025, "segment": "Total Company", "metric": "revenue", "value": 1_669_802_000.0, "period_type": "quarter"},
            {"quarter": q4_2025, "segment": "Abercrombie", "metric": "revenue", "value": 806_502_000.0, "period_type": "quarter"},
            {"quarter": q4_2025, "segment": "Hollister", "metric": "revenue", "value": 863_300_000.0, "period_type": "quarter"},
        ]
    )

    out = _filter_anf_quarterly_segment_actual_rows(
        slides_segments,
        history_revenue_by_quarter={
            q4_2024: 1_584_917_000.0,
            q4_2025: 1_669_802_000.0,
        },
    )

    kept_2024 = set(out[out["quarter"].dt.date.eq(q4_2024)]["segment"])
    kept_2025 = set(out[out["quarter"].dt.date.eq(q4_2025)]["segment"])
    assert kept_2024 == {"Americas", "EMEA", "APAC"}
    assert kept_2025 == {"Total Company", "Abercrombie", "Hollister"}


def test_anf_total_company_quarterly_revenue_falls_back_to_history_after_annual_leak_drop() -> None:
    q4_2024 = dt.date(2025, 2, 1)
    metrics = {
        "Revenue": {
            "Americas": {pd.Timestamp(q4_2024): 1_319_720_000.0},
            "EMEA": {pd.Timestamp(q4_2024): 224_467_000.0},
            "APAC": {pd.Timestamp(q4_2024): 40_730_000.0},
        }
    }

    out = _anf_add_total_company_quarter_revenue_from_history(
        metrics,
        {q4_2024: 1_584_917_000.0},
        [q4_2024],
    )

    assert out["Revenue"]["Total Company"][pd.Timestamp(q4_2024)] == pytest.approx(1_584_917_000.0)


def test_slides_guidance_has_explicit_metric_for_anf_outlook() -> None:
    slides_guidance = pd.DataFrame(
        [
            {"quarter": "2026-01-31", "period_label": "FY2026", "metric_hint": "Revenue", "numbers": "3%, 5%", "doc": "q4.htm"},
            {"quarter": "2026-01-31", "period_label": "Q1 FY2026", "metric_hint": "Adj EPS", "numbers": "$1.20, $1.30", "doc": "q4.htm"},
        ]
    )

    assert _slides_guidance_has_explicit_metric(slides_guidance, dt.date(2026, 1, 31), "Revenue", require_range=True)
    assert _slides_guidance_has_explicit_metric(slides_guidance, dt.date(2026, 1, 31), "Adj EPS", require_range=True)


def test_anf_financial_schedule_support_doc_is_not_hardcoded_to_q4_2025() -> None:
    qd = dt.date(2026, 5, 2)
    adj_metrics = pd.DataFrame(
        [
            {
                "quarter": pd.Timestamp(qd),
                "source_type": "earnings_financial_schedule",
                "doc": r"C:\Users\Jibbe\Aktier\ANF\earnings_presentation\ANF_Q1_2026_earnings_presentation_financial_schedules.pdf",
            }
        ]
    )

    doc = _anf_financial_schedule_support_doc_for_quarter(qd, adj_metrics=adj_metrics, non_gaap_files=pd.DataFrame(), slides_segments=pd.DataFrame())

    assert "ANF_Q1_2026_earnings_presentation_financial_schedules.pdf" in doc


def test_anf_visible_quarter_label_uses_retail_fiscal_year() -> None:
    assert _anf_visible_quarter_label(dt.date(2026, 1, 31)) == "2025-Q4"
    assert _anf_visible_quarter_label(dt.date(2025, 5, 3)) == "2025-Q1"
    assert _anf_visible_quarter_label(dt.date(2025, 8, 2)) == "2025-Q2"
    assert _anf_visible_quarter_label(dt.date(2025, 11, 1)) == "2025-Q3"


def test_anf_visible_ui_text_strips_badges_and_quarter_fy_labels() -> None:
    assert _anf_clean_visible_ui_text("[NEW] Q4 FY2025 brand momentum") == "2025-Q4 brand momentum"
    assert _anf_clean_visible_ui_text("[DROPPED] Dropped theme: Margin driver") == ""
    assert _anf_clean_visible_ui_text("Q1 FY2026 guidance embeds FY2026 tariff pressure") == "2026-Q1 guidance embeds 2026 year tariff pressure"


def test_anf_visible_quarter_note_summaries_split_long_multi_topic_note() -> None:
    note = (
        "[NEW] Net income per diluted share was above our outlook at $3.68 compared to $3.57 last year. "
        "We ended the quarter with inventory at cost up 5% with approximately 3 points related to tariffs. "
        "Inventory units were also up 5%, including approximately 3 points related to strategically building receipts ahead of our planned ERP implementation this month. "
        "For the year, we delivered net sales growth of 6%, reaching a record $5.27 billion. "
        "Regional growth was 7% Americas, 6% EMEA and 5% APAC."
    )

    out = _anf_visible_quarter_note_summaries(note)

    assert out == [
            "EPS / outlook: Q4 diluted EPS was $3.68, above outlook and up from $3.57 last year.",
            "Inventory: inventory cost was up 5%, including about 3 pts from tariffs.",
            "Inventory units: units were up 5%, including about 3 pts from ERP prebuild.",
            "2025 sales: net sales grew 6% to $5.27bn.",
            "Regions: 2025 year net sales grew 7% Americas, 6% EMEA and 5% APAC.",
        ]


def test_anf_driver_labels_groups_and_values_are_visible_ui_clean() -> None:
    assert _anf_compact_driver_label("Abercrombie comparable sales (%)", "%") == "Abercrombie comp"
    assert _anf_compact_driver_label("Q1 FY2026 Tariff Headwind Bps (bps)", "bps") == "2026-Q1 Tariff Headwind Bps"
    assert _anf_compact_driver_label("Total Company Inventory cost growth (%)", "%") == "Inventory cost YoY"
    assert _anf_compact_driver_group("Demand / brand momentum", "Americas comparable sales (%)", "") == "Comps"
    assert _anf_compact_driver_group("FY2026 margin bridge", "Q1 FY2026 Tariff Headwind Bps", "") == "2026 outlook bridge"
    assert _anf_round_visible_driver_value(7.000000000000001, "%", "Hollister comp", "") == pytest.approx(7.0)
    assert _anf_round_visible_driver_value(60.98058120034354, "%", "Gross margin", "") == pytest.approx(61.0)
    assert _anf_round_visible_driver_value(290.42, "bps", "Tariff headwind", "") == pytest.approx(290.0)
    assert ANF_SEGMENT_BRAND_EXPLANATION.startswith("Americas / EMEA / APAC")


def test_anf_guidance_visible_period_labels_avoid_fy_in_ui() -> None:
    assert _anf_guidance_visible_period_label("FY2026", dt.date(2026, 1, 31)) == "2026 year"
    assert _anf_guidance_visible_period_label("Q1 FY2026", dt.date(2026, 1, 31)) == "2026-Q1"
    assert _anf_guidance_visible_period_label("FY2025", dt.date(2026, 1, 31)) == "2025 year"


def test_anf_guidance_normalized_frame_filters_noise_and_cleans_horizons() -> None:
    guidance = pd.DataFrame(
        [
            {
                "quarter": "2026-01-31",
                "period_label": "FY2026",
                "period_type": "annual",
                "line": "FY2026 Revenue 3%, 5%",
                "numbers": "3%, 5%",
                "metric_hint": "Revenue",
                "low": 3,
                "high": 5,
                "value": None,
                "unit": "%",
                "doc": "ANF_2026_q4_earnings_release.htm",
                "source": "earnings_release",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "FY2026",
                "period_type": "annual",
                "line": "FY2026 Revenue around 45 million shares",
                "numbers": "45 million shares",
                "metric_hint": "Revenue",
                "low": None,
                "high": None,
                "value": 45,
                "unit": "m shares",
                "doc": "ANF_2026_q4_earnings_release.htm",
                "source": "earnings_release",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "FY2026",
                "period_type": "annual",
                "line": "FY2026 Revenue approximately 90",
                "numbers": "approximately 90",
                "metric_hint": "Revenue",
                "low": None,
                "high": None,
                "value": 90,
                "unit": None,
                "doc": "ANF_2026_q4_earnings_release.htm",
                "source": "earnings_release",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "Q1 FY2026",
                "period_type": "quarter",
                "line": "Q1 FY2026 Adj EPS $1.20, $1.30",
                "numbers": "$1.20, $1.30",
                "metric_hint": "Adj EPS",
                "low": 1.2,
                "high": 1.3,
                "value": None,
                "unit": None,
                "doc": "ANF_2026_q4_earnings_release.htm",
                "source": "earnings_release",
            },
        ]
    )

    out = _anf_visible_guidance_normalized_frame(guidance)

    assert len(out) == 2
    assert "_source_period_label" not in out.columns
    assert set(out["period_label"]) == {"2026 year", "2026-Q1"}
    assert not out["line"].astype(str).str.contains("FY2026|Q1 FY2026", regex=True).any()
    assert out[out["metric_hint"].eq("Revenue")].iloc[0]["unit"] == "%"


def test_anf_visible_guidance_normalized_frame_rejects_stale_prior_period_rows() -> None:
    guidance = pd.DataFrame(
        [
            {
                "quarter": "2025-08-02",
                "period_label": "Q1 FY2025",
                "period_type": "quarter",
                "line": "Q1 FY2025 Revenue 14%, 15%",
                "numbers": "14%, 15%",
                "metric_hint": "Revenue",
                "low": 14,
                "high": 15,
                "unit": "%",
                "doc": "ANF_Q2_2025_earnings_release.htm",
                "source": "earnings_release",
            },
            {
                "quarter": "2026-01-31",
                "period_label": "FY2025",
                "period_type": "annual",
                "line": "Fiscal 2025 outlook net sales growth at least 6%",
                "numbers": "6%",
                "metric_hint": "Revenue",
                "value": 6,
                "unit": "%",
                "doc": "ANF_Q4_2025_earnings_release.htm",
                "source": "earnings_release",
            },
            {
                "quarter": "2026-01-12",
                "period_label": "FY2025",
                "period_type": "annual",
                "line": "Business update: currently expects fiscal 2025 net sales growth at least 6%",
                "numbers": "6%",
                "metric_hint": "Revenue",
                "value": 6,
                "unit": "%",
                "doc": "ANF_January_2026_business_update.htm",
                "source": "press_release",
            },
        ]
    )

    out = _anf_visible_guidance_normalized_frame(guidance)

    assert len(out) == 1
    assert out.iloc[0]["period_label"] == "2025 year"
    assert "business update" in out.iloc[0]["line"].lower()


def test_anf_guidance_normalized_frame_cleans_forward_bridge_line_labels() -> None:
    guidance = pd.DataFrame(
        [
            {
                "quarter": "2026-01-31",
                "period_label": "Q1 FY2025",
                "period_type": "quarter",
                "line": "Q1 FY2025 tariff impact approximately 290 bps",
                "numbers": "approximately 290 bps",
                "metric_hint": "Tariff headwind",
                "low": None,
                "high": None,
                "value": 290,
                "unit": "bps",
                "doc": "8-K_2026-03-04_q4_earnings_release.htm",
                "source": "earnings_release",
            }
        ]
    )

    out = _anf_visible_guidance_normalized_frame(guidance)

    assert list(out["period_label"]) == ["2026-Q1"]
    assert "2026-Q1 tariff impact" in out.iloc[0]["line"]
    assert "2025-Q1" not in out.iloc[0]["line"]


def test_anf_promise_progress_sections_are_clean_and_open_for_2026() -> None:
    guidance = pd.DataFrame(
        [
            {"quarter": "2025-02-01", "period_label": "FY2025", "period_type": "annual", "metric_hint": "Revenue", "low": 3, "high": 5, "value": None, "unit": "%", "doc": "8-K_2025-03-06_earnings_release.htm"},
            {"quarter": "2025-05-03", "period_label": "FY2025", "period_type": "annual", "metric_hint": "Revenue", "low": 3, "high": 6, "value": None, "unit": "%", "doc": "8-K_2025-05-29_earnings_release.htm"},
            {"quarter": "2026-01-31", "period_label": "FY2026", "period_type": "annual", "metric_hint": "Revenue", "low": 3, "high": 5, "value": None, "unit": "%", "doc": "8-K_2026-03-04_earnings_release.htm"},
        ]
    )

    sections = _anf_build_promise_progress_sections(guidance, pd.DataFrame())
    progression = sections["2025 guidance progression"]
    open_rows = sections["2026 open guidance"]

    assert sum(1 for r in progression if r["Metric"] == "Net sales growth") == 1
    sales_row = next(r for r in progression if r["Metric"] == "Net sales growth")
    assert sales_row["Initial guide"] == "+3-5%"
    assert sales_row["Q1 update"] == "+3-6%"
    assert sales_row["Q2 update"] == "+5-7%"
    assert sales_row["Q3 update"] == "+6-7%"
    assert sales_row["Jan 2026 update"] == "at least +6%"
    eps_row = next(r for r in progression if r["Metric"] == "Adjusted EPS / EPS")
    assert eps_row["Initial guide"] == "$10.40-$11.40"
    assert eps_row["Q1 update"] == "$9.50-$10.50"
    assert all(r["Status"] == "Open" for r in open_rows)
    assert any(r["Horizon"] == "2026 year" and r["Metric"] == "Net sales growth" for r in open_rows)
    timeline = sections["Quarterly guidance timeline / revision log"]
    assert len(timeline) >= 8
    assert {r["Stated in"] for r in timeline} >= {"2025-Q1", "2025-Q2", "2025-Q3", "Jan 2026 pre-release update", "2025-Q4"}
    stated_order = [r["Stated in"] for r in timeline]
    assert stated_order[0] == "2025-Q4"
    assert stated_order.index("2025-Q4") < stated_order.index("Jan 2026 pre-release update") < stated_order.index("2025-Q3") < stated_order.index("2025-Q2") < stated_order.index("2025-Q1")
    jan_rows = [r for r in timeline if r["Stated in"] == "Jan 2026 pre-release update"]
    assert jan_rows
    assert all(r["Horizon"] == "2025 year" for r in jan_rows)
    assert all("before 2025 actual report" in r["Source / note"] for r in jan_rows)
    assert any(r["Metric"] == "Net sales growth" and r["Previous guide"] == "+3-6%" and r["New/current guide"] == "+5-7%" for r in timeline)
    assert all("FY" not in " ".join(str(v) for v in r.values()) for r in timeline)


def test_anf_visible_guidance_normalized_adds_stated_and_horizon_labels() -> None:
    guidance = pd.DataFrame(
        [
            {
                "quarter": "2026-01-31",
                "period_label": "FY2026",
                "period_type": "annual",
                "line": "Fiscal 2026 outlook Revenue 3%, 5%",
                "numbers": "3%, 5%",
                "metric_hint": "Revenue",
                "low": 3,
                "high": 5,
                "value": None,
                "unit": "%",
                "doc": "8-K_2026-03-04_earnings_release.htm",
                "source": "earnings_release",
            }
        ]
    )

    out = _anf_visible_guidance_normalized_frame(guidance)

    assert {"stated_in_label", "horizon_label", "horizon_type", "source_context"}.issubset(out.columns)
    assert out.iloc[0]["stated_in_label"] == "2025-Q4"
    assert out.iloc[0]["horizon_label"] == "2026 year"
    assert out.iloc[0]["period_label"] == "2026 year"


def test_anf_operating_driver_cleanup_splits_stores_tariffs_and_buybacks() -> None:
    q = dt.date(2026, 1, 31)
    rows = [
        {"Quarter": q, "_driver_key": "store_count_end", "Driver group": "Stores / real estate", "Driver": "Abercrombie stores end", "Value": 4, "Unit": "stores", "Commentary": "table fragment"},
        {"Quarter": q, "_driver_key": "store_count_end", "Driver group": "Stores / real estate", "Driver": "Company-owned stores end", "Value": 829, "Unit": "stores", "Commentary": "ended the year with 829 stores"},
        {"Quarter": q, "_driver_key": "new_stores", "Driver group": "Stores / real estate", "Driver": "Openings", "Value": 62, "Unit": "stores", "Commentary": "62 openings"},
        {"Quarter": q, "_driver_key": "closed_stores", "Driver group": "Stores / real estate", "Driver": "Closures", "Value": 22, "Unit": "stores", "Commentary": "22 closures"},
        {"Quarter": q, "_driver_key": "franchise_stores", "Driver group": "Stores / real estate", "Driver": "Franchise stores", "Value": 60, "Unit": "stores", "Commentary": "60 franchise stores"},
        {"Quarter": q, "_driver_key": "fy2026_tariff_headwind_bps", "Driver group": "FY2026 margin bridge", "Driver": "Tariff cost", "Value": 70, "Unit": "bps", "Commentary": "2026 year tariff impact 70 bps"},
        {"Quarter": q, "_driver_key": "fy2026_tariff_headwind", "Driver group": "FY2026 margin bridge", "Driver": "Tariff cost", "Value": 40, "Unit": "$m", "Commentary": "2026 year tariff cost $40m"},
        {"Quarter": q, "_driver_key": "q1_fy2026_tariff_headwind_bps", "Driver group": "FY2026 margin bridge", "Driver": "Tariff cost", "Value": 290, "Unit": "bps", "Commentary": "Q1 2026 tariff impact 290 bps"},
        {"Quarter": q, "_driver_key": "freight_tailwind_bps", "Driver group": "FY2026 margin bridge", "Driver": "Freight Tailwind Bps", "Value": 160, "Unit": "bps", "Commentary": "Q1 2026 freight tailwind 160 bps"},
        {"Quarter": q, "_driver_key": "inventory_unit_growth_erp_points", "Driver group": "Inventory / working capital", "Driver": "ERP prebuild pts", "Value": 3, "Unit": "pts", "Commentary": "ERP prebuild 3 pts"},
        {"Quarter": q, "_driver_key": "share_repurchases", "Driver group": "Capital allocation", "Driver": "Buybacks", "Value": 450, "Unit": "$m", "Commentary": "2025 repurchased $450m"},
        {"Quarter": q, "_driver_key": "old_buybacks_series", "Driver group": "Other", "Driver": "Actual buybacks", "Value": 395, "Unit": "$m", "Commentary": "mixed old row"},
        {"Quarter": q, "_driver_key": "remaining_buyback_authorization", "Driver group": "Capital allocation", "Driver": "Buybacks", "Value": 850, "Unit": "$m", "Commentary": "$850m remaining authorization"},
    ]

    out = _anf_clean_visible_operating_driver_records(rows)
    labels_list = [str(r.get("Driver")) for r in out]
    labels = set(labels_list)
    bad_store_rows = [r for r in out if "stores end" in str(r.get("Driver")).lower() and pd.to_numeric(r.get("Value"), errors="coerce") in {4, 5}]

    assert not bad_store_rows
    assert {"Owned stores end", "Actual openings", "Actual closures", "Franchise stores actual"}.issubset(labels)
    assert any(label.endswith("tariff headwind bps") for label in labels)
    assert any(label.endswith("tariff cost $m") for label in labels)
    assert {"Actual buybacks", "Guided buybacks", "Remaining authorization"}.issubset(labels)
    assert labels_list.count("Tariff cost") == 0
    assert labels_list.count("Actual buybacks") == 1
    assert labels_list.count("ERP prebuild pts") == 1
    assert "Freight Tailwind Bps" not in labels
    assert {"Sales guide", "Margin durability", "Inventory quality", "Capital returns"}.issubset(labels)
    assert all(str(r.get("Driver group")) == "Watchlist" for r in out if str(r.get("Driver")) in {"Sales guide", "Margin durability", "Inventory quality", "Capital returns"})
    by_label = {str(r.get("Driver")): r for r in out}
    assert by_label["Actual buybacks"]["Value"] == 450.0
    assert by_label["Buyback % shares"]["Value"] == 11.0
    assert by_label["Shares repurchased"]["Value"] == 5.4
    assert by_label["Avg buyback price"]["Value"] == pytest.approx(83.33, abs=0.01)


def test_anf_operating_driver_cleanup_dedupes_exact_net_sales_rows() -> None:
    q = dt.date(2026, 1, 31)
    rows = [
        {"Quarter": q, "_driver_key": "net_sales", "Driver group": "Results", "Driver": "Net sales", "Value": 1669.8, "Unit": "$m"},
        {"Quarter": q, "_driver_key": "net_sales", "Driver group": "Results", "Driver": "Net sales", "Value": 1669.8, "Unit": "$m"},
        {"Quarter": q, "_driver_key": "net_sales_other_source", "Driver group": "Results", "Driver": "Net sales", "Value": 1669.8, "Unit": "$m"},
    ]

    out = _anf_clean_visible_operating_driver_records(rows)
    compact_labels = [_anf_compact_driver_label(r.get("Driver"), r.get("Unit")) for r in out]

    assert compact_labels.count("Net sales") == 1


def test_anf_quarter_note_visible_field_polish_removes_debug_labels() -> None:
    category, metric = _anf_polish_quarter_note_visible_fields(
        "Results / drivers / better vs prior",
        "net revenue | stated Q1 2025->Q4 2025",
        "Abercrombie returned to growth and Hollister delivered its 11th consecutive quarter of growth.",
    )

    assert category == "Brand / demand"
    assert metric == "Brand momentum"
    assert "stated" not in metric.lower()

    assert _anf_visible_quarter_note_summaries("Additional lower-priority notes remain in Quarter_Notes raw sheet.") == []
    cleaned = _anf_clean_visible_ui_text("Inventory improved... while guidance remained open… and markdown risk stayed contained.", max_chars=250)
    assert "..." not in cleaned
    assert "…" not in cleaned
    assert len(cleaned) <= 250

    category, metric = _anf_polish_quarter_note_visible_fields(
        "Results / drivers / better vs prior",
        "Revenue TTM YoY at 2024-02-03",
        "Revenue TTM YoY at 2024-02-03: 15.8% (TTM $4,280.7m, LY $3,697.8m).",
    )
    assert category == "Results / financials"
    assert metric == "Results trend"

    category, metric = _anf_polish_quarter_note_visible_fields(
        "Debt / liquidity / balance sheet",
        "Net debt at 2024-02-03",
        "Net debt at 2024-02-03: $-678.8m; QoQ delta $-277.3m, YoY delta $-458.0m.",
    )
    assert category == "Capital allocation"
    assert metric == "Net cash / debt"


def test_anf_recent_operating_commentary_rows_cover_recent_quarters_without_fy_labels() -> None:
    quarters = pd.to_datetime(
        ["2024-05-04", "2024-08-03", "2024-11-02", "2025-02-01", "2025-05-03", "2025-08-02", "2025-11-01", "2026-01-31"]
    )
    hist = pd.DataFrame(
        {
            "quarter": quarters,
            "revenue": [1000, 1100, 1200, 1500, 1090, 1200, 1290, 1670],
            "gross_profit": [650, 700, 780, 930, 680, 755, 807, 993],
            "op_income": [120, 170, 180, 250, 100, 205, 155, 236],
        }
    )
    slides = pd.DataFrame(
        [
            {"quarter": "2026-01-31", "segment": "Total Company", "metric": "comparable_sales", "value": 0.01},
            {"quarter": "2026-01-31", "segment": "Abercrombie", "metric": "comparable_sales", "value": -0.01},
            {"quarter": "2026-01-31", "segment": "Hollister", "metric": "comparable_sales", "value": 0.03},
        ]
    )

    rows = _anf_recent_operating_commentary_rows(hist, slides, [q.date() for q in quarters])

    stated = {r["stated_in"] for r in rows}
    assert {"2024-Q1", "2024-Q2", "2024-Q3", "2024-Q4", "2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"}.issubset(stated)
    assert all("FY" not in r["stated_in"] for r in rows)


def test_validate_quarter_notes_accepts_legacy_evidence_column_payload() -> None:
    quarter_end = dt.date(2026, 1, 31)
    quarter_notes = pd.DataFrame(
        [
            {
                "quarter": quarter_end,
                "note_id": "history-backed",
                "claim": "Revenue was supported by History_Q at 2026-01-31.",
                "body": "Revenue was supported by History_Q at 2026-01-31.",
                "metric_ref": "revenue",
                "metric_value": 100.0,
                "evidence": '[{"doc_path":"History_Q","section_or_page":"2026-01-31","snippet":"Revenue 100.0m at 2026-01-31"}]',
                "evidence_json": "",
            }
        ]
    )
    hist = pd.DataFrame({"quarter": [quarter_end], "revenue": [100.0]})

    checks = validate_quarter_notes(quarter_notes, hist)

    assert "quarter_note_evidence_missing" not in set(checks.get("check", []))


def test_validate_quarter_notes_accepts_anf_legacy_doc_quote_evidence() -> None:
    quarter_end = dt.date(2026, 1, 31)
    quarter_notes = pd.DataFrame(
        [
            {
                "quarter": quarter_end,
                "note_id": "anf-source-note",
                "claim": "Q4 2025 brand momentum was explicit in source materials.",
                "body": "Q4 2025 brand momentum was explicit in source materials.",
                "metric_ref": "brand_family_momentum",
                "metric_value": None,
                "evidence": (
                    '{"source_type":"transcript","doc":"C:/ANF/ANF_Q4_2025_transcript.txt",'
                    '"quote":"Q4 2025 brand momentum was explicit in source materials.",'
                    '"period":"2026-01-31"}'
                ),
                "evidence_json": "",
            }
        ]
    )
    hist = pd.DataFrame({"quarter": [quarter_end], "revenue": [100.0]})

    checks = validate_quarter_notes(quarter_notes, hist)

    assert "quarter_note_evidence_incomplete" not in set(checks.get("check", []))
    assert "quarter_note_metric_nan" not in set(checks.get("check", []))


def test_annual_segment_qa_year_uses_anf_fiscal_year_for_january_year_end() -> None:
    annual_revenue_values = {
        "Americas": {2025: 4_290_395_000.0},
        "EMEA": {2025: 818_140_000.0},
        "APAC": {2025: 157_757_000.0},
    }

    assert _annual_segment_latest_year_for_qa(
        annual_revenue_values,
        dt.date(2026, 1, 31),
        is_anf_profile=True,
    ) == 2025


def test_anf_investment_case_sheet_order_places_visible_and_data_sheets() -> None:
    desired, raw_cluster = _anf_investment_case_sheet_order(
        (
            "SUMMARY",
            "Valuation",
            "BS_Segments",
            "Operating_Drivers",
            "Quarter_Notes_UI",
            "Promise_Progress_UI",
        ),
        ("History_Q", "operating_drivers_raw", "economics_market_raw"),
        is_anf_profile=True,
    )

    assert desired.index("Operating_Drivers") < desired.index("ANF_Investment_Case") < desired.index("Quarter_Notes_UI")
    assert "ANF_Investment_Case_Data" in raw_cluster
    assert raw_cluster.index("ANF_Investment_Case_Data") > raw_cluster.index("operating_drivers_raw")


def test_sector_investment_case_sheet_order_places_all_supported_tickers_after_operating_drivers() -> None:
    base_desired = (
        "SUMMARY",
        "Valuation",
        "BS_Segments",
        "Operating_Drivers",
        "Economics_Overlay",
        "Quarter_Notes_UI",
        "Promise_Progress_UI",
    )
    base_raw = ("History_Q", "operating_drivers_raw", "economics_market_raw")

    for ticker in ("ANF", "PBI", "GPRE"):
        desired, raw_cluster = _investment_case_sheet_order(base_desired, base_raw, ticker=ticker)
        case_sheet = f"{ticker}_Investment_Case"
        data_sheet = f"{ticker}_Investment_Case_Data"
        assert desired[desired.index("Operating_Drivers") + 1] == case_sheet
        assert desired.index(case_sheet) < desired.index("Quarter_Notes_UI")
        if ticker == "GPRE":
            assert desired.index("Quarter_Notes_UI") < desired.index("Economics_Overlay")
        assert data_sheet in raw_cluster
        assert raw_cluster.index(data_sheet) == raw_cluster.index("operating_drivers_raw") + 1


def test_sector_investment_case_data_is_sector_specific_without_cross_contamination() -> None:
    hist = pd.DataFrame(
        [
            {"quarter": dt.date(2025, 3, 31), "revenue": 900_000_000.0, "ebitda": 90_000_000.0, "cfo": 80_000_000.0, "capex": 25_000_000.0, "cash": 200_000_000.0, "total_debt": 1_100_000_000.0, "shares_diluted": 100_000_000.0},
            {"quarter": dt.date(2025, 6, 30), "revenue": 910_000_000.0, "ebitda": 95_000_000.0, "cfo": 82_000_000.0, "capex": 25_000_000.0, "cash": 205_000_000.0, "total_debt": 1_080_000_000.0, "shares_diluted": 100_000_000.0},
            {"quarter": dt.date(2025, 9, 30), "revenue": 920_000_000.0, "ebitda": 100_000_000.0, "cfo": 85_000_000.0, "capex": 26_000_000.0, "cash": 210_000_000.0, "total_debt": 1_050_000_000.0, "shares_diluted": 100_000_000.0},
            {"quarter": dt.date(2025, 12, 31), "revenue": 930_000_000.0, "ebitda": 105_000_000.0, "cfo": 88_000_000.0, "capex": 27_000_000.0, "cash": 215_000_000.0, "total_debt": 1_020_000_000.0, "shares_diluted": 100_000_000.0},
        ]
    )

    pbi = _sector_build_investment_case_data(ticker="PBI", hist=hist)
    gpre = _sector_build_investment_case_data(ticker="GPRE", hist=hist)
    pbi_text = " ".join(str(x) for x in pbi.to_numpy().ravel()).lower()
    gpre_text = " ".join(str(x) for x in gpre.to_numpy().ravel()).lower()

    assert "presort" in pbi_text
    assert "sendtech" in pbi_text
    assert "refinancing" in pbi_text
    assert "ethanol" not in pbi_text
    assert "hollister" not in pbi_text
    assert "abercrombie" not in pbi_text

    assert "ethanol" in gpre_text
    assert "45z" in gpre_text
    assert "crush margin" in gpre_text
    assert "rin" in gpre_text
    assert "presort" not in gpre_text
    assert "sendtech" not in gpre_text
    assert "hollister" not in gpre_text

    assert set(pbi["section"]).issuperset(
        {
            "Investment Snapshot",
            "What needs to happen for the stock to work",
            "Turnaround / EBIT Bridge",
            "FCF / Debt Paydown Bridge",
            "Buybacks vs FCF",
            "Current Guide -> Implied Earnings",
            "What Moves EPS",
            "Segment Health",
            "Segment Trend / Lapping Risk",
            "Capital Structure / Refinancing Risk",
            "Guidance Beat/Miss Setup",
            "Valuation Sensitivity",
            "Adj EBITDA x EV/EBITDA",
            "FCF Yield Implied Equity Value",
        }
    )
    assert "Segment / Business Health" not in set(pbi["section"])
    assert set(gpre["section"]).issuperset(
        {
            "Investment Snapshot",
            "What needs to happen for the stock to work",
            "Ethanol / Crush Margin Bridge",
            "Policy / 45Z / RFS Bridge",
            "Buybacks vs FCF",
            "Current Guide -> Implied Earnings",
            "What Moves EPS",
            "What Moves EBITDA",
            "Margin Cycle / Lapping Risk",
            "Ethanol / Policy Health",
            "FCF / Balance Sheet",
            "Guidance Beat/Miss Setup",
            "Valuation Sensitivity",
            "Adj EBITDA x EV/EBITDA",
            "FCF Yield Implied Equity Value",
        }
    )


def test_sector_investment_case_writer_creates_readable_visible_and_audit_sheets() -> None:
    hist = pd.DataFrame(
        [
            {"quarter": dt.date(2025, 3, 31), "revenue": 100_000_000.0, "ebitda": 10_000_000.0, "net_income": 3_000_000.0, "shares_diluted": 10_000_000.0, "cfo": 8_000_000.0, "capex": 2_000_000.0, "cash": 20_000_000.0, "total_debt": 50_000_000.0},
            {"quarter": dt.date(2025, 6, 30), "revenue": 110_000_000.0, "ebitda": 11_000_000.0, "net_income": 3_500_000.0, "shares_diluted": 10_000_000.0, "cfo": 9_000_000.0, "capex": 2_000_000.0, "cash": 21_000_000.0, "total_debt": 49_000_000.0},
            {"quarter": dt.date(2025, 9, 30), "revenue": 120_000_000.0, "ebitda": 12_000_000.0, "net_income": 4_000_000.0, "shares_diluted": 10_000_000.0, "cfo": 10_000_000.0, "capex": 2_000_000.0, "cash": 22_000_000.0, "total_debt": 48_000_000.0},
            {"quarter": dt.date(2025, 12, 31), "revenue": 130_000_000.0, "ebitda": 13_000_000.0, "net_income": 4_500_000.0, "shares_diluted": 10_000_000.0, "cfo": 11_000_000.0, "capex": 2_000_000.0, "cash": 23_000_000.0, "total_debt": 47_000_000.0},
        ]
    )

    for ticker in ("PBI", "GPRE"):
        data = _sector_build_investment_case_data(ticker=ticker, hist=hist)
        wb = Workbook()
        del wb[wb.sheetnames[0]]
        _write_sector_investment_case_sheet(wb, ticker, data)
        _write_sector_investment_case_data_sheet(wb, ticker, data)

        ws = wb[f"{ticker}_Investment_Case"]
        assert ws.sheet_view.zoomScale >= 110
        assert ws.freeze_panes == "A4"
        assert ws.cell(1, 1).value == f"{ticker} Investment Case"
        assert ws.cell(1, 1).font.sz == 16
        section_rows = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value in set(data["section"])]
        assert section_rows
        for row in section_rows:
            assert ws.row_dimensions[row].height == 21
            assert ws.cell(row, 1).font.sz == 12
        assert ws.cell(section_rows[0] + 2, 1).font.sz == 12
        assert 40 <= ws.column_dimensions["A"].width <= 44
        assert ws.column_dimensions["B"].width >= 24
        assert ws.column_dimensions["D"].width >= 24
        assert ws.column_dimensions["E"].width >= 24
        snapshot_rows = range(section_rows[0] + 1, section_rows[0] + 8)
        assert all(float(ws.row_dimensions[r].height or 0.0) == pytest.approx(21.0, abs=0.1) for r in snapshot_rows)
        top_values = [
            str(ws.cell(r, c).value or "").strip()
            for r in range(1, min(ws.max_row, 20) + 1)
            for c in range(1, min(ws.max_column, 10) + 1)
        ]
        assert "Topic" not in top_values
        assert not any(ws.cell(r, c).value == "Model read" for r in range(1, 21) for c in range(2, 11))
        assert ws.cell(section_rows[0] + 1, 1).value == "Model read"
        snapshot_merged = [str(rng) for rng in ws.merged_cells.ranges if str(rng).startswith(f"B{section_rows[0] + 1}:")]
        assert snapshot_merged and snapshot_merged[0].endswith(f"J{section_rows[0] + 1}")
        assert any(str(rng).startswith("B") and ":G" in str(rng) for rng in ws.merged_cells.ranges)
        rendered_sections = {
            str(ws.cell(r, 1).value or "").strip()
            for r in range(1, ws.max_row + 1)
            if str(ws.cell(r, 1).value or "").strip()
        }
        assert {
            "Buybacks vs FCF",
            "Current Guide -> Implied Earnings",
            "What Moves EPS",
            "Valuation Sensitivity",
            "Adj EBITDA x EV/EBITDA",
            "FCF Yield Implied Equity Value",
            "Guidance Beat/Miss Setup",
        }.issubset(rendered_sections)
        if ticker == "PBI":
            assert {"Segment Trend / Lapping Risk", "Segment Health"}.issubset(rendered_sections)
            assert "Segment / Business Health" not in rendered_sections
        else:
            assert {"Margin Cycle / Lapping Risk", "Ethanol / Policy Health"}.issubset(rendered_sections)
        adj_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Adj EBITDA x EV/EBITDA")
        assert [ws.cell(adj_row + 1, c).value for c in range(1, 6)] == [
            "Multiple",
            "EV",
            "Equity value: core net cash",
            "Share price",
            "Source / investment read",
        ]
        valuation_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Valuation Sensitivity")
        assert [ws.cell(valuation_row + 1, c).value for c in range(1, 6)] == ["EPS", "10x", "12x", "14x", "16x"]
        scenario_row = next(r for r in range(valuation_row + 1, ws.max_row + 1) if ws.cell(r, 1).value == "Scenario")
        assert [ws.cell(scenario_row, c).value for c in range(1, 5)] == ["Scenario", "EPS", "P/E", "Share price"]
        fcf_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "FCF Yield Implied Equity Value")
        assert [ws.cell(fcf_row + 1, c).value for c in range(1, 5)] == [
            "Yield",
            "Equity value",
            "Share price",
            "Source / note",
        ]
        long_right_aligned = [
            cell.coordinate
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=10)
            for cell in row
            if isinstance(cell.value, str)
            and len(cell.value) > 25
            and cell.alignment.horizontal == "right"
        ]
        assert not long_right_aligned
        assert f"{ticker}_Investment_Case_Data" in wb.sheetnames


def test_gpre_source_backed_debt_tranche_fallback_dedupes_current_schedule() -> None:
    slides_debt = pd.DataFrame(
        [
            {"quarter": "2026-03-31", "tranche": "2.25% convertible notes due 2027 (1)", "amount": 60_000_000, "maturity_year": 2027, "is_table_total": False, "asof_match_found": True, "doc": "source.htm", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "5.25% convertible notes due 2030 (2)", "amount": 200_000_000, "maturity_year": 2030, "is_table_total": False, "asof_match_found": True, "doc": "source.htm", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "Term loan due 2035 (3)", "amount": 69_750_000, "maturity_year": 2035, "is_table_total": False, "asof_match_found": True, "doc": "source.htm", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "Green Plains Central City Carbon Capture Tallgrass Term loan due 2038", "amount": 44_126_000, "maturity_year": 2038, "is_table_total": False, "asof_match_found": True, "doc": "source.htm", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "Green Plains Wood River Carbon Capture Tallgrass Term loan due 2038", "amount": 48_387_000, "maturity_year": 2038, "is_table_total": False, "asof_match_found": True, "doc": "source.htm", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "Green Plains York Carbon Capture Tallgrass Term loan due 2037", "amount": 34_389_000, "maturity_year": 2037, "is_table_total": False, "asof_match_found": True, "doc": "source.htm", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "Other", "amount": 9_661_000, "maturity_year": None, "is_table_total": False, "asof_match_found": True, "doc": "source.htm", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "2.25% convertible notes due 2027 (1) $ 60,000 $ 60,000", "amount": 60_000_000, "maturity_year": 2027, "is_table_total": False, "asof_match_found": True, "doc": "source.pdf", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "2.25% convertible notes due 2027 1,897 —", "amount": 1_897_000, "maturity_year": 2027, "is_table_total": False, "asof_match_found": False, "doc": "source.pdf", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "5.25% convertible notes due 2030 12,723 —", "amount": 12_723_000, "maturity_year": 2030, "is_table_total": False, "asof_match_found": False, "doc": "source.pdf", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "Tallgrass Term loan due 2038 44,126 —", "amount": 44_126_000, "maturity_year": 2038, "is_table_total": False, "asof_match_found": True, "doc": "source.pdf", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "Tallgrass Term loan due 2038 48,387 —", "amount": 48_387_000, "maturity_year": 2038, "is_table_total": False, "asof_match_found": True, "doc": "source.pdf", "source": "financial_statement"},
            {"quarter": "2026-03-31", "tranche": "Principal amount", "amount": 466_313_000, "maturity_year": None, "is_table_total": True, "asof_match_found": True, "doc": "source.htm", "source": "financial_statement"},
        ]
    )

    out = _source_backed_debt_tranches_from_slides(slides_debt, "2026-03-31")

    assert len(out) == 7
    assert out["amount_principal"].sum() == pytest.approx(466_313_000.0)
    values = dict(zip(out["tranche_name"], out["amount_principal"]))
    assert values["2.25% convertible notes due 2027 (1)"] == pytest.approx(60_000_000.0)
    assert values["5.25% convertible notes due 2030 (2)"] == pytest.approx(200_000_000.0)
    assert values["Term loan due 2035 (3)"] == pytest.approx(69_750_000.0)
    assert values["Green Plains Central City Carbon Capture Tallgrass Term loan due 2038"] == pytest.approx(44_126_000.0)
    assert values["Green Plains Wood River Carbon Capture Tallgrass Term loan due 2038"] == pytest.approx(48_387_000.0)
    assert values["Green Plains York Carbon Capture Tallgrass Term loan due 2037"] == pytest.approx(34_389_000.0)
    assert values["Other"] == pytest.approx(9_661_000.0)
    near_term = out.loc[out["tranche_name"].eq("2.25% convertible notes due 2027 (1)"), "near_term"].iloc[0]
    assert bool(near_term)
    assert out["source_basis"].astype(str).str.contains("within 24 months of latest quarter end", regex=False).all()


def test_shared_promise_progress_postprocess_dedupes_all_timeline_metrics() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Promise_Progress_UI"
    section_fill = PatternFill("solid", fgColor="5B9BD5")
    ws.cell(1, 1, "Promise Progress")
    ws.cell(3, 1, "2025-Q1 revisions").fill = section_fill
    headers = [
        "Metric",
        "Previous guide",
        "New/current guide",
        "Change type",
        "Actual / latest actual",
        "Status",
        "Horizon",
        "Stated in",
        "Source date",
        "Source / note",
    ]
    for cc, header in enumerate(headers, start=1):
        ws.cell(4, cc, header)
    duplicate = [
        "Cost savings target",
        "$180m-$200m",
        "$180m-$200m",
        "Maintained",
        "$157m run-rate",
        "On track",
        "2025-Q1",
        "2025-Q1",
        "2025-06-30",
        "confirmed",
    ]
    for rr, source_date in enumerate(["2025-06-30", "2025-09-30", "2025-12-31"], start=5):
        row = list(duplicate)
        row[8] = source_date
        for cc, value in enumerate(row, start=1):
            ws.cell(rr, cc, value)
    changed = [
        "Cost savings target",
        "$150m-$170m",
        "$180m-$200m",
        "Updated",
        "$157m run-rate",
        "On track",
        "2025-Q1",
        "2025-Q1",
        "2025-03-31",
        "raised",
    ]
    for cc, value in enumerate(changed, start=1):
        ws.cell(8, cc, value)

    _apply_shared_ui_conventions_to_workbook(wb, "PBI")

    rows = [
        [ws.cell(rr, cc).value for cc in range(1, 11)]
        for rr in range(1, ws.max_row + 1)
    ]
    cost_rows = [row for row in rows if row[0] == "Cost savings target"]
    assert len(cost_rows) == 1
    assert cost_rows[0][1] == "$150m-$170m"
    assert cost_rows[0][2] == "$180m-$200m"
    assert cost_rows[0][3] == "Updated"


def test_shared_promise_progress_postprocess_removes_semantic_carry_forward_rows() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Promise_Progress_UI"
    section_fill = PatternFill("solid", fgColor="5B9BD5")
    headers = [
        "Metric",
        "Previous guide",
        "New/current guide",
        "Change type",
        "Actual / latest actual",
        "Status",
        "Horizon",
        "Stated in",
        "Source date",
        "Source / note",
    ]

    ws.cell(1, 1, "Promise Progress")
    ws.cell(3, 1, "2025-Q4 revisions").fill = section_fill
    for cc, header in enumerate(headers, start=1):
        ws.cell(4, cc, header)
    gpre_rows = [
        [
            "Capex guidance (2026 year)",
            "$15.0m-$25.0m",
            "$15.0m-$25.0m",
            "Maintained",
            "not yet measurable",
            "Open",
            "2026 year",
            "2025-Q4",
            "2026-03-31",
            "Later Q1 carry-forward confirmation; should not live inside 2025-Q4 revisions.",
        ],
        [
            "Capex guidance (2026 year)",
            "",
            "$15.0m-$25.0m",
            "Initial",
            "not yet measurable",
            "Open",
            "2026 year",
            "2025-Q4",
            "2025-12-31",
            "Initial 2025-Q4 sustaining capex guidance.",
        ],
    ]
    for rr, row in enumerate(gpre_rows, start=5):
        for cc, value in enumerate(row, start=1):
            ws.cell(rr, cc, value)

    ws.cell(9, 1, "2025-Q1 revisions").fill = section_fill
    for cc, header in enumerate(headers, start=1):
        ws.cell(10, cc, header)
    pbi_rows = [
        [
            "Cost savings target",
            "$150m-$170m",
            "$180m-$200m",
            "Raised",
            "not yet measurable",
            "Open",
            "2025 year",
            "2025-Q1",
            "2025-03-31",
            "True Q1 target increase.",
        ],
        [
            "Cost savings target",
            "$180m-$200m",
            "$180m-$200m",
            "Maintained",
            "$157m run-rate",
            "On track",
            "2025 year",
            "2025-Q1",
            "2026-03-31",
            "Later carry-forward confirmation; latest run-rate belongs in current guidance section.",
        ],
    ]
    for rr, row in enumerate(pbi_rows, start=11):
        for cc, value in enumerate(row, start=1):
            ws.cell(rr, cc, value)

    _apply_shared_ui_conventions_to_workbook(wb, "PBI")

    rows = [
        [ws.cell(rr, cc).value for cc in range(1, 11)]
        for rr in range(1, ws.max_row + 1)
    ]
    capex_rows = [row for row in rows if row[0] == "Capex guidance (2026 year)"]
    assert len(capex_rows) == 1
    assert capex_rows[0][3] == "Initial"
    assert capex_rows[0][8] == "2025-12-31"

    cost_rows = [row for row in rows if row[0] == "Cost savings target"]
    assert len(cost_rows) == 1
    assert cost_rows[0][1] == "$150m-$170m"
    assert cost_rows[0][2] == "$180m-$200m"
    assert cost_rows[0][3] == "Raised"
    assert cost_rows[0][8] == "2025-03-31"
    assert "$157m run-rate" not in {str(row[4]) for row in cost_rows}


def test_shared_promise_progress_postprocess_preserves_true_initial_rows() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Promise_Progress_UI"
    section_fill = PatternFill("solid", fgColor="5B9BD5")
    ws.cell(1, 1, "Promise Progress")
    ws.cell(3, 1, "2024-Q2 revisions").fill = section_fill
    headers = [
        "Metric",
        "Previous guide",
        "New/current guide",
        "Change type",
        "Actual / latest actual",
        "Status",
        "Horizon",
        "Stated in",
        "Source date",
        "Source / note",
    ]
    for cc, header in enumerate(headers, start=1):
        ws.cell(4, cc, header)
    row = [
        "Cost savings target",
        "",
        "$75m-$85m",
        "Initial",
        "$157m run-rate",
        "Open",
        "2024 year",
        "2024-Q2",
        "2024-06-30",
        "Raised target to $180m-$200m annualized savings; latest disclosed $157m run-rate.",
    ]
    for cc, value in enumerate(row, start=1):
        ws.cell(5, cc, value)

    _apply_shared_ui_conventions_to_workbook(wb, "PBI")

    out = [ws.cell(5, cc).value for cc in range(1, 11)]
    assert out[0] == "Cost savings target"
    assert out[1] in {"", None}
    assert out[2] == "$75m-$85m"
    assert out[3] == "Initial"
    assert out[4] == "not yet measurable"
    assert "180m-$200m" not in str(out[9])
    assert "run-rate" not in str(out[9]).lower()


def test_shared_promise_progress_postprocess_evaluates_completed_horizons_with_actuals() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Promise_Progress_UI"
    section_fill = PatternFill("solid", fgColor="5B9BD5")
    ws.cell(1, 1, "Promise Progress")
    ws.cell(3, 1, "2025-Q2 revisions").fill = section_fill
    headers = [
        "Metric",
        "Previous guide",
        "New/current guide",
        "Change type",
        "Actual / latest actual",
        "Status",
        "Horizon",
        "Stated in",
        "Source date",
        "Source / note",
    ]
    for cc, header in enumerate(headers, start=1):
        ws.cell(4, cc, header)
    rows = [
        ["FCF target", "$330m-$370m", "$330m-$370m", "Maintained", "$383.3m", "On track", "2025 year", "2025-Q2", "2025-06-30", "Final 2025 actual shown for evaluation."],
        ["Capex", "~$245m", "~$245m", "Maintained", "$240.8m", "On track", "2025 year", "2025-Q2", "2025-06-30", "Final 2025 actual shown for evaluation."],
        ["Adjusted EPS / EPS", "$10.20-$10.50", "$10.20-$10.50", "Maintained", "$10.46 GAAP / $9.86 adjusted", "On track", "2025 year", "2025-Q2", "2025-06-30", "Basis differs."],
    ]
    for rr, row in enumerate(rows, start=5):
        for cc, value in enumerate(row, start=1):
            ws.cell(rr, cc, value)

    _apply_shared_ui_conventions_to_workbook(wb, "ANF")

    out = [[ws.cell(rr, cc).value for cc in range(1, 11)] for rr in range(5, 8)]
    assert out[0][5] == "Beat"
    assert out[1][5] == "Hit"
    assert out[2][5] == "Basis-dependent"


def test_shared_promise_progress_postprocess_does_not_reclassify_expected_milestones() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Promise_Progress_UI"
    section_fill = PatternFill("solid", fgColor="5B9BD5")
    ws.cell(1, 1, "Promise Progress")
    ws.cell(3, 1, "2025-Q3 revisions").fill = section_fill
    headers = [
        "Metric",
        "Previous guide",
        "New/current guide",
        "Change type",
        "Actual / latest actual",
        "Status",
        "Horizon",
        "Stated in",
        "Source date",
        "Source / note",
    ]
    for cc, header in enumerate(headers, start=1):
        ws.cell(4, cc, header)
    row = [
        "45Z facility qualification",
        "",
        "All 8 plants qualified for 45Z",
        "Initial",
        "Expected in 2026",
        "On track",
        "2025-Q3",
        "2025-Q3",
        "2025-09-30",
        "All eight plants expected to qualify for 45Z in 2026.",
    ]
    for cc, value in enumerate(row, start=1):
        ws.cell(5, cc, value)

    _apply_shared_ui_conventions_to_workbook(wb, "GPRE")

    assert ws.cell(5, 6).value == "On track"


def test_pbi_gpre_guidance_normalized_cleanup_preserves_qualitative_guidance() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Guidance_Normalized"
    headers = ["source_date", "stated_in_label", "horizon_label", "metric", "value", "low", "high", "unit", "basis", "line", "source_context", "source"]
    for cc, header in enumerate(headers, start=1):
        ws.cell(1, cc, header)
    rows = [
        ["2026-03-31", "2026-Q1", "2026 year", "45Z facility qualification", "", "", "", "", "qualitative", "Management expects remaining facilities to qualify for 45Z during 2026.", "expects remaining facilities to qualify for 45Z during 2026", "earnings_release"],
        ["2014-03-31", "2014-Q1", "2014 year", "Adj EPS", "", "0.42", "15", "$", "", "year; updates GAAP EPS from continuing operations guidance Issued $500 million of 10-year bonds and retired $500 million of debt", "year; updates GAAP EPS from continuing operations guidance Issued $500 million of 10-year bonds and retired $500 million of debt", "financial_statement"],
        ["2024-03-31", "2024-Q1", "2024 year", "Revenue", "", "50", "100", "%", "", "to repeat sales customers and anticipate expanding it from there with the goal of eventually moving to 100% of our production", "to repeat sales customers and anticipate expanding it", "transcript"],
        ["2019-09-30", "2019-Q3", "2019 year", "Revenue", "", "1", "8217", "$m", "", "Interim period results are not necessarily indicative of results for the full year", "Interim period results are not necessarily indicative", "financial_statement"],
        ["2018-03-31", "2018-Q1", "2018 year", "Revenue", "", "1", "2", "$m", "", "The Company adopted ASC Topic 606 revenue recognition guidance.", "ASC Topic 606 revenue recognition guidance", "financial_statement"],
        ["2019-03-31", "2019-Q1", "2019 year", "Lease accounting", "", "1", "2", "$m", "", "The Company implemented a lease accounting system related to ASC Topic 842.", "ASC Topic 842 lease accounting system", "financial_statement"],
        ["2025-03-31", "2025-Q1", "2025 year", "Revenue", "", "1", "2", "$m", "", "Good morning and thank you for joining today's call.", "Good morning and thank you", "webcast transcript"],
    ]
    for rr, row in enumerate(rows, start=2):
        for cc, value in enumerate(row, start=1):
            ws.cell(rr, cc, value)

    _apply_shared_ui_conventions_to_workbook(wb, "GPRE")

    remaining = [
        " | ".join(str(ws.cell(rr, cc).value or "") for cc in range(1, ws.max_column + 1))
        for rr in range(2, ws.max_row + 1)
    ]
    assert len(remaining) == 1
    assert "45Z facility qualification" in remaining[0]
    assert "Issued $500 million" not in remaining[0]
    assert "repeat sales customers" not in remaining[0]
    assert "Interim period results" not in remaining[0]
    assert "ASC Topic 606" not in remaining[0]
    assert "ASC Topic 842" not in remaining[0]
    assert "Good morning and thank you" not in remaining[0]

def test_pbi_needs_review_date_columns_are_wide_enough_for_display() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Needs_Review"
    headers = ["priority", "issue_family", "severity", "first_seen_q", "last_seen_q", "latest_message", "recommended_action"]
    for cc, header in enumerate(headers, start=1):
        ws.cell(1, cc, header)
    ws.append(["P2", "qsum_vs_fy", "warn", dt.datetime(2025, 12, 31), dt.datetime(2026, 3, 31), "Sum of 4 quarters vs FY fact differs.", "Review as fiscal/YTD limitation."])

    _apply_shared_ui_conventions_to_workbook(wb, "PBI")

    assert ws.column_dimensions["D"].width >= 18
    assert ws.column_dimensions["E"].width >= 18
    assert ws.column_dimensions["F"].width >= 48
    assert ws.column_dimensions["G"].width >= 50
    assert ws["D2"].number_format == "yyyy-mm-dd"
    assert ws["E2"].number_format == "yyyy-mm-dd"
    assert str(ws["C2"].value).lower() == "warn"


def test_shared_promise_progress_rewrite_puts_values_before_metadata_and_newest_first() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Promise_Progress_UI"
    rows = [
        ["Promise Progress"],
        ["Generated at 2026-05-15 00:00:00 UTC | Quarter blocks"],
        ["Promise progress (As of 2026-03-31)"],
        ["Metric", "Target", "Latest", "Result", "Rationale", "Stated", "Last Seen", "Carried To", "Evaluated Through", "Evidence"],
        ["Revenue guidance", "$1.8bn-$1.86bn", "not yet measurable", "Open", "2026 year Revenue guidance $1.8bn-$1.86bn.", "2026-Q1", "2026-Q1", "2026-Q1", "2026-03-31", ""],
        ["Adjusted EBIT guidance", "$425m-$465m", "not yet measurable", "Open", "2026 year Adjusted EBIT guidance $425m-$465m.", "2026-Q1", "2026-Q1", "2026-Q1", "2026-03-31", ""],
        ["Promise progress (As of 2025-12-31)"],
        ["Metric", "Target", "Latest", "Result", "Rationale", "Stated", "Last Seen", "Carried To", "Evaluated Through", "Evidence"],
        ["Revenue guidance", "$1.76bn-$1.86bn", "not yet measurable", "Open", "2026 year Revenue guidance $1.76bn-$1.86bn.", "2025-Q4", "2025-Q4", "2026-Q1", "2026-03-31", ""],
        ["Promise progress (As of 2025-09-30)"],
        ["Metric", "Target", "Latest", "Result", "Rationale", "Stated", "Last Seen", "Carried To", "Evaluated Through", "Evidence"],
        ["Revenue actual", "$1.8bn", "1892629000.0", "Completed", "Actual revenue was 1892629000.0.", "2025-Q3", "2025-Q3", "2025-Q3", "2025-09-30", ""],
        ["Adj EPS actual", "$1.36", "1.36", "Completed", "Actual adjusted EPS was 1.36.", "2025-Q3", "2025-Q3", "2025-Q3", "2025-09-30", ""],
    ]
    for rr, row in enumerate(rows, start=1):
        for cc, value in enumerate(row, start=1):
            ws.cell(rr, cc, value)

    _rewrite_shared_promise_progress_ui_from_blocks(ws, ticker="PBI")

    values = [[ws.cell(rr, cc).value for cc in range(1, 11)] for rr in range(1, ws.max_row + 1)]
    assert any(row[0] == "Current guidance progression" for row in values)
    assert any(row[0] == "Open guidance" for row in values)
    timeline_header = next(row for row in values if row[:10] == [
        "Metric",
        "Previous guide",
        "New/current guide",
        "Change type",
        "Actual / latest actual",
        "Status",
        "Horizon",
        "Stated in",
        "Source date",
        "Source / note",
    ])
    assert timeline_header[0] == "Metric"
    first_revenue = next(row for row in values if row[0] == "Revenue guidance" and row[8] == "2026-03-31")
    assert first_revenue[2] == "$1.8bn-$1.86bn"
    assert first_revenue[5] == "Open"
    assert first_revenue[7] == "2026-Q1"
    timeline_metric_rows = [row for row in values if row[0] in {"Revenue guidance", "Adjusted EBIT guidance"} and row[8]]
    assert timeline_metric_rows[0][7] == "2026-Q1"
    assert timeline_metric_rows[-1][7] == "2025-Q4"
    formatted_actual = next(row for row in values if row[0] == "Revenue actual" and row[8] == "2025-09-30")
    assert formatted_actual[4] == "$1.89bn"
    formatted_eps = next(row for row in values if row[0] == "Adj EPS actual" and row[8] == "2025-09-30")
    assert formatted_eps[4] == "$1.36"
    assert ws.freeze_panes in {"A2", None}
    def _has_merge(row_idx: int, start_col: int, end_col: int) -> bool:
        return any(
            rng.min_row == row_idx
            and rng.max_row == row_idx
            and rng.min_col == start_col
            and rng.max_col == end_col
            for rng in ws.merged_cells.ranges
        )

    current_section_row = next(rr for rr in range(1, ws.max_row + 1) if ws.cell(rr, 1).value == "Current guidance progression")
    open_section_row = next(rr for rr in range(1, ws.max_row + 1) if ws.cell(rr, 1).value == "Open guidance")
    assert _has_merge(current_section_row + 1, 6, 10)
    assert _has_merge(current_section_row + 2, 6, 10)
    assert _has_merge(open_section_row + 1, 5, 10)
    assert _has_merge(open_section_row + 2, 5, 10)


def test_shared_promise_progress_rewrite_compacts_long_guidance_value_cells() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Promise_Progress_UI"
    rows = [
        ["Promise Progress"],
        ["Promise progress (As of 2026-03-31)"],
        ["Metric", "Target", "Latest", "Result", "Rationale", "Stated", "Last Seen", "Carried To", "Evaluated Through", "Evidence"],
        [
            "45Z facility qualification",
            "All 8 plants qualified for 45Z tax credits",
            "All 8 plants qualified/expected to qualify in 2026",
            "Completed",
            "All eight operating plants qualified/expected to qualify for 45Z tax credits in 2026; on-farm practice upside remains excluded pending final Treasury guidance/calculator.",
            "2026-Q1",
            "2026-Q1",
            "2026-Q1",
            "2026-03-31",
            "",
        ],
        [
            "Advantage Nebraska EBITDA opportunity",
            "$140m-$165m in 2026",
            "Advantage Nebraska fully operational and sequestering CO2 in Wyoming",
            "On track",
            "Advantage Nebraska expected to contribute $140m-$165m to 2026 year 45Z EBITDA.",
            "2026-Q1",
            "2026-Q1",
            "2026-Q1",
            "2026-03-31",
            "",
        ],
    ]
    for rr, row in enumerate(rows, start=1):
        for cc, value in enumerate(row, start=1):
            ws.cell(rr, cc, value)

    _rewrite_shared_promise_progress_ui_from_blocks(ws, ticker="GPRE")

    values = [[ws.cell(rr, cc).value for cc in range(1, 11)] for rr in range(1, ws.max_row + 1)]
    facility_row = next(row for row in values if row[0] == "45Z facility qualification" and row[8] == "2026-03-31")
    assert facility_row[2] == "All 8 qualified"
    assert facility_row[4] == "All 8 qualified"
    assert len(facility_row[2]) < 22
    assert len(facility_row[4]) < 22
    nebraska_row = next(row for row in values if row[0] == "Advantage Nebraska EBITDA opportunity" and row[8] == "2026-03-31")
    assert nebraska_row[4] == "AN operational"


def test_shared_promise_progress_rewrite_keeps_future_annual_guidance_open_or_on_track() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Promise_Progress_UI"
    rows = [
        ["Promise Progress"],
        ["Promise progress (As of 2026-03-31)"],
        ["Metric", "Target", "Latest", "Result", "Rationale", "Stated", "Last Seen", "Carried To", "Evaluated Through", "Evidence"],
        [
            "45Z EBITDA guidance",
            "$200m-$225m",
            "$55.2m Q1 realized",
            "Completed",
            "2026 year 45Z EBITDA guidance has Q1 actual but annual horizon is not complete.",
            "2026-Q1",
            "2026-Q1",
            "2026-Q1",
            "2026-03-31",
            "",
        ],
    ]
    for rr, row in enumerate(rows, start=1):
        for cc, value in enumerate(row, start=1):
            ws.cell(rr, cc, value)

    _rewrite_shared_promise_progress_ui_from_blocks(ws, ticker="GPRE")

    values = [[ws.cell(rr, cc).value for cc in range(1, 11)] for rr in range(1, ws.max_row + 1)]
    annual_row = next(row for row in values if row[0] == "45Z EBITDA guidance" and row[6] == "2026 year")
    assert annual_row[5] in {"On track", "Open"}
    assert annual_row[5] != "Completed"


def test_shared_ui_polish_cleans_guidance_normalized_and_downgrades_pbi_qsum() -> None:
    wb = Workbook()
    wb.active.title = "Valuation"
    guidance = wb.create_sheet("Guidance_Normalized")
    guidance.append(["source_date", "stated_in_label", "horizon_label", "metric", "value", "source_note"])
    guidance.append(["2026-03-31", "2026-Q1", "2026 year", "Revenue", "$1.8bn-$1.86bn", "management guidance"])
    guidance.append(["2026-03-31", "2026-Q1", "", "Other", "noise", "raw Other row"])
    guidance.append(["2025-12-31", "2025-Q4", "2026 year", "Revenue", "$1.8bn", "&#8220; raw html fragment"])
    guidance.append(["2026-03-31", "2026-Q1", "2026 year", "Capex", "$100m", "Investing activities were primarily affected by lower capital expenditures"])
    guidance.append(["2026-03-31", "2026-Q1", "", "Revenue", "$477m", r"C:\PBI_raised_fy2026_guidance.pdf actual result compared to prior year"])
    guidance.append(["2026-03-31", "2026-Q1", "2026 year", "Adj EPS", "$0.42 to 15", "impossible EPS range"])
    guidance.append(["2026-03-31", "2026-Q1", "2026 year", "Revenue", "$0.40 to $0.01", "unit mismatch"])
    guidance.append(["2026-03-31", "2026-Q1", "2026 year", "Revenue", "8 to 3", "ambiguous unitless range"])
    guidance.append(["2026-03-31", "2026-Q1", "2026 year", "Other", "Competition Alternative fuels", "analyst question fragment?"])
    guidance.append(["2026-03-31", "2026-Q1", "2026 year", "Revenue", "$900m", "Actual revenue was $900m compared to prior year"])
    guidance.append(["2026-03-31", "2026-Q1", "2026 year", "Revenue", "1%, 1, 1, 2", "Good morning and thank you for taking the questions. Can you comment on subscribers?"])
    guidance.append(["2026-03-31", "2026-Q1", "2029", "Debt", "$500m", "ABL facility matures August 2029; no guidance or outlook."])
    guidance.append(
        [
            "2026-03-31",
            "2026-Q1",
            "2027",
            "Debt",
            "$60m",
            "Outside of the $60 million of 2027 convertible notes that remain outstanding, we anticipate retiring the notes with cash at maturity.",
        ]
    )
    guidance.append(
        [
            "2026-03-31",
            "2026-Q1",
            "2026 year",
            "Other",
            "7.2",
            "Stock Options The fair value of the stock options is estimated on the date of the grant using the Black-Scholes model.",
        ]
    )
    needs = wb.create_sheet("Needs_Review")
    needs.append(["priority", "severity", "check", "latest_message", "first_seen", "last_seen", "recommended_action"])
    needs.column_dimensions["E"].width = 8
    needs.column_dimensions["F"].width = 8
    needs.append(["P1", "fail", "qsum_vs_fy", "Sum of 4 quarters vs FY fact.", "2026-03-31", "2026-05-17", ""])

    _apply_shared_ui_conventions_to_workbook(wb, ticker="PBI")

    remaining = [
        [guidance.cell(rr, cc).value for cc in range(1, guidance.max_column + 1)]
        for rr in range(2, guidance.max_row + 1)
    ]
    assert remaining == [["2026-03-31", "2026-Q1", "2026 year", "Revenue", "$1.8bn-$1.86bn", "management guidance"]]
    assert needs["A2"].value == "P2"
    assert needs["B2"].value == "warn"
    assert "older fiscal/YTD" in str(needs["G2"].value)
    assert needs.column_dimensions["E"].width >= 13
    assert needs.column_dimensions["F"].width >= 13
    assert needs["E2"].number_format == "yyyy-mm-dd"


def test_shared_ui_polish_standardizes_promise_statuses_and_valuation_row_heights() -> None:
    wb = Workbook()
    val = wb.active
    val.title = "Valuation"
    for rr in range(1, 30):
        val.cell(rr, 1, "Revenue" if rr >= 8 else "")
        val.row_dimensions[rr].height = 22.0
    val["A3"] = "Valuation"
    val["A6"] = "Quarter"
    val["A7"] = "Operating"
    val["O7"] = "Guidance (As of 2026-01-31) - Status: Open | Found: Revenue"
    val["O7"].fill = PatternFill("solid", fgColor="5B9BD5")
    val["O8"] = "Metric"
    val["O9"] = "Revenue"
    val["O10"] = "Cash flow"
    val["O10"].fill = PatternFill("solid", fgColor="5B9BD5")
    promise = wb.create_sheet("Promise_Progress_UI")
    promise.append(["Metric", "Previous guide", "Latest guide", "Status"])
    promise.append(["Sales growth", "", "+6%", "met"])
    promise.append(["Operating margin", "", "13%", "met-ish"])
    promise.append(["EPS", "", "$10.30-$10.40", "basis-dependent"])
    promise.append(["Capex", "", "$245m", "mixed"])
    promise.append(["EPS basis", "", "", "", "", "", "basis-dependent"])
    promise.append([])
    promise.append(["Metric", "Previous guide", "New/current guide", "Change type", "Actual / latest actual", "Status"])
    promise.append(["Revenue", "+3-5%", "+4-6%", "on track", "$1.2bn", "open"])

    _apply_shared_ui_conventions_to_workbook(wb, ticker="ANF")

    assert val.row_dimensions[9].height == 19.5
    assert val.row_dimensions[10].height == 19.5
    assert val.row_dimensions[20].height == 19.5
    assert val.row_dimensions[7].height == 19.5
    assert "Found:" not in str(val["O7"].value)
    assert val["A3"].font.sz == 18
    assert [promise.cell(rr, 4).value for rr in range(2, 6)] == [
        "Completed",
        "On track",
        "Basis-dependent",
        "Mixed",
    ]
    assert promise["G6"].value == "Basis-dependent"
    fills = {promise.cell(rr, 4).fill.fgColor.rgb for rr in range(2, 6)}
    assert len(fills) == 4
    timeline_row = 9
    assert promise.cell(timeline_row, 4).value == "Updated"
    assert promise.cell(timeline_row, 6).value == "Open"
    assert promise.cell(timeline_row, 4).fill.fgColor.rgb != promise.cell(timeline_row, 6).fill.fgColor.rgb
    assert promise.cell(timeline_row, 5).fill.fgColor.rgb != promise.cell(timeline_row, 6).fill.fgColor.rgb


def test_shared_quarter_notes_category_polish_is_sector_specific_and_removes_debug_markers() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quarter_Notes_UI"
    ws.append(["Quarter", "Category", "Note", "Metric"])
    ws.append(["2025-Q4", "Programs / initiatives", "Cost savings flowed through...", "Cost savings"])
    ws.append(["", "Debt / refi / covenants", "[NEW] Liquidity improved\u2026", "Debt"])
    _standardize_quarter_notes_ui_categories(ws, ticker="PBI")
    assert ws["B2"].value == "Cost savings"
    assert ws["B3"].value == "Balance sheet / liquidity"
    assert "..." not in str(ws["C2"].value)
    assert "\u2026" not in str(ws["C3"].value)
    assert "[NEW]" not in str(ws["C3"].value)

    wb_g = Workbook()
    ws_g = wb_g.active
    ws_g.title = "Quarter_Notes_UI"
    ws_g.append(["Quarter", "Category", "Note", "Metric"])
    ws_g.append(["2025-Q4", "guidance", "45Z guidance and RVO policy supported the setup.", "45Z"])
    ws_g.append(["", "production", "Produced gallons increased.", "Produced gallons"])
    _standardize_quarter_notes_ui_categories(ws_g, ticker="GPRE")
    assert ws_g["B2"].value == "45Z / carbon"
    assert ws_g["B3"].value == "Production / gallons"


def test_sector_operating_driver_intro_tables_are_ordered_and_sector_specific() -> None:
    pbi_tables = _sector_operating_driver_intro_tables("PBI")
    gpre_tables = _sector_operating_driver_intro_tables("GPRE")
    anf_tables = _sector_operating_driver_intro_tables("ANF")

    assert [tbl["title"] for tbl in pbi_tables[:2]] == ["Current watchlist", "Current/latest outlook"]
    assert [tbl["title"] for tbl in gpre_tables[:2]] == ["Current watchlist", "Current/latest outlook"]
    assert [tbl["title"] for tbl in anf_tables[:2]] == ["Current watchlist", "Current/latest outlook"]

    pbi_text = " ".join(str(v) for tbl in pbi_tables for row in tbl["rows"] for v in row).lower()
    gpre_text = " ".join(str(v) for tbl in gpre_tables for row in tbl["rows"] for v in row).lower()
    anf_text = " ".join(str(v) for tbl in anf_tables for row in tbl["rows"] for v in row).lower()

    assert "presort" in pbi_text and "sendtech" in pbi_text and "refinancing" in pbi_text
    assert "ethanol" not in pbi_text and "hollister" not in pbi_text
    assert "crush" in gpre_text and "45z" in gpre_text and "e15" in gpre_text
    assert "presort" not in gpre_text and "hollister" not in gpre_text
    assert "comps" in anf_text and "tariff" in anf_text and "stores" in anf_text
    assert "ethanol" not in anf_text and "presort" not in anf_text


def test_anf_investment_case_data_contains_bridges_sensitivities_and_retail_sections() -> None:
    hist = pd.DataFrame(
        [
            {
                "quarter": dt.date(2026, 1, 31),
                "revenue": 5_266_292_000.0,
                "op_income": 699_143_000.0,
                "net_income": 506_921_000.0,
                "diluted_shares": 48.476,
                "inventory": 601_218_000.0,
            }
        ]
    )
    driver_rows = [
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "q1_fy2026_tariff_headwind_bps", "Driver group": "FY2026 margin bridge", "Driver": "Q1 tariff", "Value": 290, "Unit": "bps"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "q1_fy2026_tariff_headwind", "Driver group": "FY2026 margin bridge", "Driver": "Q1 tariff cost", "Value": 30, "Unit": "$m"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "fy2026_tariff_headwind_bps", "Driver group": "FY2026 margin bridge", "Driver": "FY tariff", "Value": 70, "Unit": "bps"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "fy2026_tariff_headwind", "Driver group": "FY2026 margin bridge", "Driver": "FY tariff cost", "Value": 40, "Unit": "$m"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "freight_tailwind_bps", "Driver group": "FY2026 margin bridge", "Driver": "Freight", "Value": 160, "Unit": "bps"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "marketing_headwind_bps", "Driver group": "FY2026 margin bridge", "Driver": "Marketing", "Value": 50, "Unit": "bps"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "inventory_cost_growth", "Driver group": "Inventory / working capital", "Driver": "Inventory cost", "Value": 0.05, "Unit": "%"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "inventory_cost_tariff_points", "Driver group": "Inventory / working capital", "Driver": "Tariff component", "Value": 3, "Unit": "pts"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "inventory_unit_growth", "Driver group": "Inventory / working capital", "Driver": "Inventory units", "Value": 0.05, "Unit": "%"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "inventory_unit_growth_erp_points", "Driver group": "Inventory / working capital", "Driver": "ERP prebuild", "Value": 3, "Unit": "pts"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "inventory_unit_growth_ex_erp", "Driver group": "Inventory / working capital", "Driver": "Inventory ex ERP", "Value": 0.02, "Unit": "%"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "store_count_end", "Driver group": "Stores / real estate", "Driver": "Owned stores end", "Value": 829, "Unit": "stores"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "franchise_stores", "Driver group": "Stores / real estate", "Driver": "Franchise stores", "Value": 60, "Unit": "stores"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "total_stores_including_franchise", "Driver group": "Stores / real estate", "Driver": "Total stores", "Value": 889, "Unit": "stores"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "new_stores", "Driver group": "Stores / real estate", "Driver": "Openings", "Value": 62, "Unit": "stores"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "closed_stores", "Driver group": "Stores / real estate", "Driver": "Closures", "Value": 22, "Unit": "stores"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "average_buyback_price", "Driver group": "Capital allocation", "Driver": "Avg buyback price", "Value": 83.33, "Unit": "$/share"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "remaining_buyback_authorization", "Driver group": "Capital allocation", "Driver": "Remaining authorization", "Value": 850, "Unit": "$m"},
        {"Quarter": dt.date(2026, 1, 31), "_driver_key": "digital_sales_mix", "Driver group": "Digital / omnichannel", "Driver": "Digital sales mix", "Value": 44, "Unit": "%"},
    ]
    slides_segments = pd.DataFrame(
        [
            {"quarter": "2025-05-03", "segment": "Total Company", "metric": "comparable_sales", "value": 0.04, "period_type": "quarter"},
            {"quarter": "2025-05-03", "segment": "Abercrombie", "metric": "comparable_sales", "value": -0.10, "period_type": "quarter"},
            {"quarter": "2025-05-03", "segment": "Hollister", "metric": "comparable_sales", "value": 0.23, "period_type": "quarter"},
            {"quarter": "2024-05-04", "segment": "Total Company", "metric": "comparable_sales", "value": 0.22, "period_type": "quarter"},
            {"quarter": "2026-01-31", "segment": "Total Company", "metric": "comparable_sales", "value": 0.01, "period_type": "quarter"},
            {"quarter": "2026-01-31", "segment": "Abercrombie", "metric": "comparable_sales", "value": -0.01, "period_type": "quarter"},
            {"quarter": "2026-01-31", "segment": "Hollister", "metric": "comparable_sales", "value": 0.03, "period_type": "quarter"},
            {"quarter": "2025-02-01", "segment": "Total Company", "metric": "comparable_sales", "value": 0.16, "period_type": "quarter"},
            {"quarter": "2026-01-31", "segment": "Abercrombie", "metric": "revenue", "value": 806_502_000.0, "period_type": "quarter"},
            {"quarter": "2026-01-31", "segment": "Hollister", "metric": "revenue", "value": 863_300_000.0, "period_type": "quarter"},
            {"quarter": "2026-01-31", "segment": "Abercrombie", "metric": "revenue", "value": 2_523_662_000.0, "period_type": "annual"},
            {"quarter": "2026-01-31", "segment": "Hollister", "metric": "revenue", "value": 2_742_630_000.0, "period_type": "annual"},
            {"quarter": "2026-01-31", "segment": "Abercrombie", "metric": "net_sales_growth", "value": -0.01, "period_type": "annual"},
            {"quarter": "2026-01-31", "segment": "Hollister", "metric": "net_sales_growth", "value": 0.15, "period_type": "annual"},
            {"quarter": "2026-01-31", "segment": "Abercrombie", "metric": "net_sales_growth", "value": 0.04, "period_type": "quarter"},
            {"quarter": "2026-01-31", "segment": "Hollister", "metric": "net_sales_growth", "value": 0.06, "period_type": "quarter"},
        ]
    )
    guidance = pd.DataFrame(
        [
            {"quarter": "2026-01-31", "period_label": "2026 year", "period_type": "annual", "metric_hint": "Revenue", "low": 3, "high": 5, "unit": "%"},
            {"quarter": "2026-01-31", "period_label": "2026 year", "period_type": "annual", "metric_hint": "Operating margin", "low": 12.0, "high": 12.5, "unit": "%"},
            {"quarter": "2026-01-31", "period_label": "2026 year", "period_type": "annual", "metric_hint": "Adj EPS", "low": 10.2, "high": 11.0, "unit": "$/share"},
            {"quarter": "2026-01-31", "period_label": "2026 year", "period_type": "annual", "metric_hint": "Share repurchases", "value": 450, "unit": "$m"},
            {"quarter": "2026-01-31", "period_label": "2026 year", "period_type": "annual", "metric_hint": "Capex", "low": 200, "high": 225, "unit": "$m"},
            {"quarter": "2026-01-31", "period_label": "2026 year", "period_type": "annual", "metric_hint": "Diluted shares", "value": 45, "unit": "m shares"},
        ]
    )

    out = _anf_build_investment_case_data(
        hist=hist,
        operating_driver_rows=driver_rows,
        guidance_normalized=guidance,
        slides_segments=slides_segments,
        valuation_summary=pd.DataFrame(),
        adjusted_metrics=pd.DataFrame(),
    )

    key = {(str(r["section"]), str(r["metric"])): r for r in out.to_dict("records")}
    assert key[("Tariff / Margin Bridge", "Q1 2026 tariff headwind")]["display"] == "~290 bps / ~$30m"
    assert key[("Tariff / Margin Bridge", "2026 tariff headwind")]["display"] == "~70 bps / ~$40m incremental"
    assert key[("EPS Bridge", "2025 adjusted EPS")]["display"] == "$9.86"
    assert key[("EPS Bridge", "Sales growth")]["display"] == "+3-5%"
    assert key[("EPS Bridge", "2026 guided EPS")]["display"] == "$10.20-$11.00"
    assert key[("Investment Snapshot", "Upside path")]["display"]
    assert key[("Investment Snapshot", "Current stance based on model data")]["display"]
    assert key[("What Needs To Happen", "Margins must stabilize")]["display"].startswith("Operating margin")
    assert key[("Buybacks vs FCF", "FCF TTM")]["display"].startswith("$")
    assert "Watch" in key[("Buybacks vs FCF", "Investment read")]["display"]
    assert key[("What Moves EPS", "+100 bps operating margin")]["value"] > 0
    assert key[("Comp Stack / Lapping Risk", "2025-Q1 2-year stack")]["display"] == "+26%"
    assert key[("Comp Stack / Lapping Risk", "2025-Q4 2-year stack")]["display"] == "+17%"
    assert key[("Comp Stack / Lapping Risk", "2025-Q4 2-year stack")]["short_read"]
    assert key[("Brand Health", "Abercrombie 2025 sales")]["display"] == "~$2.52bn"
    assert key[("Brand Health", "Hollister 2025 sales")]["display"] == "~$2.74bn"
    assert key[("Inventory / Markdown Risk", "Inventory cost tariff component")]["display"] == "~3 pts"
    assert key[("Store Productivity / Real Estate ROI", "Company-owned stores")]["value"] == 829
    assert key[("Store Productivity / Real Estate ROI", "2026 remodels/right-sizes")]["value"] == 70
    assert key[("Store Productivity / Real Estate ROI", "Method note")]["display"].startswith("Sales/store is a proxy")
    assert "year-end owned stores" in key[("Store Productivity / Real Estate ROI", "Sales per owned store")]["source_note"]
    assert key[("Guidance Beat/Miss Setup", "Sales growth")]["display"] == "+3-5%"


def test_anf_investment_case_data_has_guide_to_implied_earnings_and_cross_checks() -> None:
    hist = pd.DataFrame(
        [
            {
                "quarter": dt.date(2026, 1, 31),
                "revenue": 5_266_292_000.0,
                "op_income": 699_143_000.0,
                "net_income": 506_921_000.0,
                "diluted_shares": 48.476,
            }
        ]
    )
    guidance = pd.DataFrame(
        [
            {"quarter": "2026-01-31", "period_label": "2026 year", "period_type": "annual", "metric_hint": "Revenue", "low": 3, "high": 5, "unit": "%"},
            {"quarter": "2026-01-31", "period_label": "2026 year", "period_type": "annual", "metric_hint": "Operating margin", "low": 12.0, "high": 12.5, "unit": "%"},
            {"quarter": "2026-01-31", "period_label": "2026 year", "period_type": "annual", "metric_hint": "Adj EPS", "low": 10.2, "high": 11.0, "unit": "$/share"},
            {"quarter": "2026-01-31", "period_label": "2026 year", "period_type": "annual", "metric_hint": "Diluted shares", "value": 45, "unit": "m shares"},
        ]
    )
    valuation = pd.DataFrame(
        [
            {"metric": "Adjusted EBITDA TTM", "value": 815.590},
            {"metric": "FCF TTM", "value": 378.368},
            {"metric": "Net cash incl. marketable securities", "value": 784.576},
            {"metric": "Lease-adjusted net debt incl. securities", "value": 383.519},
        ]
    )

    out = _anf_build_investment_case_data(
        hist=hist,
        operating_driver_rows=[],
        guidance_normalized=guidance,
        slides_segments=pd.DataFrame(),
        valuation_summary=valuation,
        adjusted_metrics=pd.DataFrame(),
    )

    key = {(str(r["section"]), str(r["metric"])): r for r in out.to_dict("records")}
    guide_section = "2026 Guide → Implied Earnings"
    assert key[(guide_section, "2025 revenue")]["display"] == "$5,266.3m"
    assert key[(guide_section, "2026 revenue growth guide")]["display"] == "+3-5%"
    assert key[(guide_section, "Operating margin guide")]["display"] == "12.0-12.5%"
    assert key[(guide_section, "Diluted shares guide")]["display"] == "45.0m"
    assert key[(guide_section, "Company EPS guide")]["display"] == "$10.20-$11.00"
    assert key[(guide_section, "Implied EPS low/high")]["value_low"] == pytest.approx(10.33, abs=0.15)
    assert key[(guide_section, "Implied EPS low/high")]["value_high"] == pytest.approx(10.98, abs=0.15)
    assert key[("Adj EBITDA x EV/EBITDA", "8.0x EV/EBITDA")]["equity_value_core_net_cash"] > 7_000
    assert key[("FCF Yield Implied Equity Value", "5.0% FCF yield")]["share_price"] > 150
    assert key[("Buybacks vs FCF", "2025 buybacks")]["display"] == "~$450m"
    assert key[("Buybacks vs FCF", "Shares repurchased")]["display"] == "5.4m"
    assert key[("Digital / omnichannel", "Digital sales mix")]["display"] == "44%"
    assert "Abercrombie" in key[("Brand Health", "2025 sales")]["display"]
    assert all(str(key[(guide_section, metric)].get("source_note") or "").strip() for metric in ["2025 revenue", "Company EPS guide", "Model vs guide check"])


def test_anf_investment_case_writer_applies_readable_style_and_formula_grid() -> None:
    data = pd.DataFrame(
        [
            {"section": "Investment Snapshot", "metric": "Model read", "display": "Constructive but margin-sensitive."},
            {"section": "Investment Snapshot", "metric": "Why it can work", "display": "Net cash, strong FCF, buyback capacity and Hollister momentum."},
            {"section": "Investment Snapshot", "metric": "Key debate", "display": "Whether tariff, ERP and marketing headwinds are temporary enough."},
            {"section": "Investment Snapshot", "metric": "What improves case", "display": "Better comps and tariff mitigation."},
            {"section": "Investment Snapshot", "metric": "What breaks case", "display": "Hollister slowdown or markdown pressure."},
            {"section": "Investment Snapshot", "metric": "Watch next", "display": "Q1 2026 comps and gross margin bridge."},
            {"section": "Key Debate", "metric": "Key debate", "display": "Can ANF sustain high margins and EPS after the turnaround?", "source_note": "decision frame"},
            {"section": "Tariff / Margin Bridge", "metric": "Q1 2026 tariff headwind", "display": "~290 bps / ~$30m", "q1_display": "~290 bps / ~$30m", "year_display": ""},
            {"section": "Tariff / Margin Bridge", "metric": "2026 tariff headwind", "display": "~70 bps / ~$40m incremental", "q1_display": "", "year_display": "~70 bps / ~$40m incremental"},
            {"section": "Tariff / Margin Bridge", "metric": "Freight tailwind", "display": "~160 bps", "q1_display": "~160 bps", "year_display": "partial annual offset"},
            {"section": "Tariff / Margin Bridge", "metric": "ERP disruption", "display": ">100 bps op margin headwind", "q1_display": ">100 bps op margin headwind", "year_display": "temporary"},
            {"section": "Tariff / Margin Bridge", "metric": "Marketing", "display": "+50 bps headwind Q1", "q1_display": "+50 bps headwind Q1", "year_display": "strategic spend"},
            {"section": "Tariff / Margin Bridge", "metric": "AUR / pricing", "display": "offset / mitigation", "q1_display": "offset / mitigation", "year_display": "partial mitigation"},
            {"section": "Tariff / Margin Bridge", "metric": "Sourcing / supplier mitigation", "display": "offset", "q1_display": "offset", "year_display": "partial mitigation"},
            {"section": "Tariff / Margin Bridge", "metric": "Reported 2025 operating margin", "display": "13.3%"},
            {"section": "Tariff / Margin Bridge", "metric": "2026 guide operating margin", "display": "12.0-12.5%"},
            {"section": "Tariff / Margin Bridge", "metric": "Implied decline", "display": "-130 to -80 bps"},
            {"section": "EPS Bridge", "metric": "2025 adjusted EPS", "display": "$9.86"},
            {"section": "EPS Bridge", "metric": "Sales growth", "display": "+3-5%"},
            {"section": "EPS Bridge", "metric": "Margin / tariff / freight / AUR", "display": "tariffs + ERP + marketing, partly offset by freight/AUR/pricing/sourcing"},
            {"section": "EPS Bridge", "metric": "SG&A leverage / deleverage", "display": "depends on sales growth and marketing investment"},
            {"section": "EPS Bridge", "metric": "Buyback / share count reduction", "display": "~45m guided diluted shares vs 48.5m in 2025"},
            {"section": "EPS Bridge", "metric": "2026 guided EPS", "display": "$10.20-$11.00"},
            {"section": "2026 Guide → Implied Earnings", "metric": "2025 revenue", "display": "$5,266.3m"},
            {"section": "2026 Guide → Implied Earnings", "metric": "2026 revenue growth guide", "display": "+3-5%"},
            {"section": "2026 Guide → Implied Earnings", "metric": "Implied 2026 revenue", "display": "$5,424.3-$5,529.6m"},
            {"section": "2026 Guide → Implied Earnings", "metric": "Operating margin guide", "display": "12.0-12.5%"},
            {"section": "2026 Guide → Implied Earnings", "metric": "Implied EBIT", "display": "$650.9-$691.2m"},
            {"section": "2026 Guide → Implied Earnings", "metric": "Tax / interest assumptions", "display": "28.5% tax / $0m net interest"},
            {"section": "2026 Guide → Implied Earnings", "metric": "Diluted shares guide", "display": "45.0m"},
            {"section": "2026 Guide → Implied Earnings", "metric": "Implied EPS low/high", "display": "$10.33-$10.98"},
            {"section": "2026 Guide → Implied Earnings", "metric": "Company EPS guide", "display": "$10.20-$11.00"},
            {"section": "2026 Guide → Implied Earnings", "metric": "Model vs guide check", "display": "Ties to guide."},
            {"section": "What Moves EPS", "metric": "+100 bps operating margin", "display": "+$0.88 EPS"},
            {"section": "What Moves EPS", "metric": "+100 bps gross margin", "display": "+$0.88 EPS before SG&A leakage"},
            {"section": "What Moves EPS", "metric": "+1% sales growth", "display": "+$0.11 EPS"},
            {"section": "What Moves EPS", "metric": "$100m buybacks", "display": "+$0.29 EPS at $83.33/share"},
            {"section": "What Moves EPS", "metric": "Roughly +$1 EPS equals", "display": "~114 bps op margin or ~$345m buybacks"},
            {"section": "Valuation Sensitivity", "metric": "EPS $9.50", "value": 9.5, "display": "$9.50"},
            {"section": "Valuation Sensitivity", "metric": "EPS $10.50", "value": 10.5, "display": "$10.50"},
            {"section": "Valuation Sensitivity", "metric": "EPS $11.50", "value": 11.5, "display": "$11.50"},
            {"section": "Valuation Sensitivity", "metric": "Base scenario", "scenario": "Base", "eps": 10.6, "multiple": 13, "share_price": 138, "display": "$138"},
            {"section": "Adj EBITDA x EV/EBITDA", "metric": "8.0x EV/EBITDA", "display": "$6,525m EV", "equity_value_core_net_cash": 7310.0, "equity_value_lease_adjusted": 6140.0},
            {"section": "FCF Yield Implied Equity Value", "metric": "5.0% FCF yield", "display": "$7,567m equity value", "share_price": 168.2},
        ]
    )
    wb = Workbook()
    del wb[wb.sheetnames[0]]

    _write_anf_investment_case_sheet(wb, data)
    _write_anf_investment_case_data_sheet(wb, data)

    ws = wb["ANF_Investment_Case"]
    assert ws.sheet_view.zoomScale >= 110
    assert ws.freeze_panes == "A4"
    assert ws.max_column >= 10
    assert ws.cell(1, 1).font.sz == 16
    assert ws.cell(4, 1).font.sz == 12
    assert ws.row_dimensions[4].height == 21
    assert ws.cell(4, 1).value == "Investment Snapshot"
    assert any(ws.cell(r, 1).value == "Key Debate" for r in range(1, ws.max_row + 1))
    assert ws.column_dimensions["B"].width >= 38
    assert ws.column_dimensions["C"].width >= 44
    assert ws.column_dimensions["D"].width >= 38
    assert ws.column_dimensions["J"].width >= 20
    tariff_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Tariff / Margin Bridge")
    tariff_header_row = tariff_row + 1
    assert ws.cell(tariff_row, 1).font.sz == 12
    assert ws.row_dimensions[tariff_row].height == 21
    assert ws.cell(tariff_header_row, 1).font.sz == 12
    assert ws.cell(tariff_header_row, 6).font.sz == 12
    merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
    assert f"B{tariff_header_row}:C{tariff_header_row}" in merged_ranges
    assert f"D{tariff_header_row}:E{tariff_header_row}" in merged_ranges
    assert f"F{tariff_header_row}:J{tariff_header_row}" in merged_ranges
    eps_header_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "EPS Bridge") + 1
    guide_header_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "2026 Guide → Implied Earnings") + 1
    assert f"D{eps_header_row}:J{eps_header_row}" in merged_ranges
    assert f"D{guide_header_row}:J{guide_header_row}" in merged_ranges
    assert any(ws.cell(r, 1).value == "2026 Guide → Implied Earnings" for r in range(1, ws.max_row + 1))
    assert any(str(ws.cell(r, c).value or "").startswith("=") for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1))
    data_ws = wb["ANF_Investment_Case_Data"]
    headers = [data_ws.cell(1, c).value for c in range(1, data_ws.max_column + 1)]
    assert "source_note" in headers


def test_sector_investment_case_writer_uses_anf_style_callout_and_wrapped_labels() -> None:
    data = pd.DataFrame(
        [
            {"section": "Investment Snapshot", "metric": "Model read", "display": "Commodity/policy upside case with margin durability risk."},
            {"section": "Investment Snapshot", "metric": "Why it can work", "display": "45Z and crush-margin support can lift EBITDA if it converts into FCF."},
            {"section": "Investment Snapshot", "metric": "Key debate", "display": "Can policy and crush-margin support become durable enough for a higher multiple?"},
            {"section": "Investment Snapshot", "metric": "What would improve the case", "display": "45Z cash conversion, exports/E15 demand and disciplined capex."},
            {"section": "Investment Snapshot", "metric": "What would break the case", "display": "Policy disappointment or crush margins normalizing lower."},
            {"section": "Investment Snapshot", "metric": "Watch next", "display": "45Z implementation, RVO/SRE/RIN policy and capex."},
            {"section": "Investment Snapshot", "metric": "Current stance", "display": "Upside needs policy value to convert into FCF."},
            {
                "section": "What needs to happen for the stock to work",
                "metric": "RVO/SRE/RIN policy support",
                "display": "Policy must support demand and margins.",
                "source": "model-derived",
            },
            {
                "section": "Ethanol / Crush Margin Bridge",
                "metric": "Crush margin proxy",
                "display": "Economics_Overlay translates ethanol/corn/coproduct spreads into EBITDA sensitivity.",
                "source": "Economics_Overlay",
            },
        ]
    )
    wb = Workbook()
    del wb[wb.sheetnames[0]]

    _write_sector_investment_case_sheet(wb, "GPRE", data)

    ws = wb["GPRE_Investment_Case"]
    top_values = [str(ws.cell(r, c).value or "").strip() for r in range(1, 21) for c in range(1, 11)]
    assert top_values.count("Topic") == 0
    assert not any(v == "Model read" and c != 1 for r in range(1, 21) for c in range(1, 11) for v in [str(ws.cell(r, c).value or "").strip()])
    assert "Model read" in [str(ws.cell(r, 1).value or "").strip() for r in range(1, 21)]
    merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
    assert "B5:J5" in merged_ranges
    key_debate_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Key Debate")
    assert ws.cell(key_debate_row + 1, 1).value == "Debate"
    assert f"B{key_debate_row + 1}:J{key_debate_row + 1}" in merged_ranges
    long_label_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "RVO/SRE/RIN policy support")
    assert ws.cell(long_label_row, 1).alignment.wrap_text is True
    assert ws.cell(long_label_row, 1).fill.fgColor.rgb != "005B9BD5"


def test_sector_investment_case_data_replaces_generic_gpre_overlay_placeholders() -> None:
    hist = pd.DataFrame(
        [
            {"quarter": "2025-06-30", "revenue": 400_000_000, "ebitda": 20_000_000, "cfo": 10_000_000, "capex": 4_000_000, "debt_core": 450_000_000, "cash": 90_000_000},
            {"quarter": "2025-09-30", "revenue": 420_000_000, "ebitda": 30_000_000, "cfo": 20_000_000, "capex": 5_000_000, "debt_core": 455_000_000, "cash": 95_000_000},
            {"quarter": "2025-12-31", "revenue": 430_000_000, "ebitda": 40_000_000, "cfo": 30_000_000, "capex": 6_000_000, "debt_core": 456_000_000, "cash": 96_000_000},
            {"quarter": "2026-03-31", "revenue": 445_800_000, "ebitda": 49_500_000, "cfo": 66_400_000, "capex": 26_900_000, "debt_core": 458_200_000, "cash": 95_700_000},
        ]
    )
    guidance = pd.DataFrame(
        [
            {
                "metric": "45Z EBITDA guidance",
                "target_display": "$200m-$225m",
                "horizon_label": "2026 year",
                "source_note": "management guidance",
            }
        ]
    )

    out = _sector_build_investment_case_data(ticker="GPRE", hist=hist, guidance_normalized=guidance)

    bridge = out[out["section"].eq("Ethanol / Crush Margin Bridge")]
    displays = [str(v or "") for v in bridge["display"].tolist()]
    assert not any(v.strip() == "See Economics_Overlay market model." for v in displays)
    assert any("EBITDA sensitivity" in v or "spread" in v.lower() for v in displays)
    policy = out[(out["section"].eq("Policy / 45Z / RFS Bridge")) & (out["metric"].eq("45Z expected benefit"))]
    assert not policy.empty
    assert "$200m-$225m" in str(policy.iloc[0]["display"])
    visible_blob = " | ".join(str(v or "") for v in out.astype(object).to_numpy().ravel())
    assert "Guidance_Normalized" not in visible_blob
    assert "Slides_Guidance / curated guidance profile" in visible_blob


def test_shared_readable_source_type_label_for_side_panels() -> None:
    assert _shared_readable_source_type_label("earnings_presentation") == "earnings presentation"
    assert _shared_readable_source_type_label("history_q") == "quarterly history"
    assert _shared_readable_source_type_label("promise") == "management guidance"
    assert _shared_readable_source_type_label("quarter_note") == "quarter notes"


def test_anf_valuation_side_panel_scrubber_replaces_strays_with_intentional_blocks() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation"
    ws["A1"] = "Valuation"
    ws["O16"] = "Guidance (As of 2025-11-01) - Status: Maintained | Found: none"
    ws["O17"] = "Metric"
    ws["O18"] = "No guidance items for this quarter."
    ws["O33"] = "Operating Drivers"
    ws["O34"] = "Driver group"
    ws["O40"] = "Thesis Bridge"
    ws["O42"] = "Bridge item"
    ws["A170"] = "Guidance detail"
    ws["B170"] = "See Promise_Progress_UI and ANF_Investment_Case for the current guidance bridge."
    ws["AA8"] = "Trend / realized"
    for rr in range(7, 45):
        ws.row_dimensions[rr].height = 24.0
    ws.merge_cells("O16:V16")
    ws.merge_cells("O40:V40")

    _anf_clear_valuation_side_panels(ws)
    _write_anf_valuation_side_panel(ws)

    visible = "\n".join(
        str(ws.cell(r, c).value or "")
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    )
    assert visible.count("Guidance (As of") == 2
    assert visible.count("Operating Drivers") == 1
    assert visible.count("Thesis Bridge") == 1
    assert "No guidance items for this quarter" not in visible
    assert "Guidance detail" not in visible
    assert "See Promise_Progress_UI" not in visible
    assert "Memo:" not in visible
    assert "Q1 FY2026" not in visible
    assert "FY2026" not in visible
    assert "2026 year" in visible
    assert "2026-Q1" in visible
    assert "+3-5%" in visible
    assert "$10.20-$11.00" in visible
    assert "$815.6m" in visible
    assert "Adj EBITDA TTM" in visible
    assert "Hollister growth engine" in visible
    assert "earnings release / quarterly history" in visible
    assert "quarterly history" in visible
    assert "annual report / transcript" in visible
    assert "rel / qtr" not in visible
    assert "qtr hist." not in visible
    assert "rel / tr." not in visible
    assert "Open / not yet realized" in visible
    assert ws["AA8"].value == "Trend / realized"

    side_panel_text = "\n".join(
        str(ws.cell(r, c).value or "")
        for r in range(1, ws.max_row + 1)
        for c in range(15, 23)
    ).lower()
    for bad in ("parcel", "mail volumes", "cost savings target", "protein", "ethanol", "crush", "rin", "45z", "debt paydown uplift"):
        assert bad not in side_panel_text

    guidance_rows = [
        r
        for r in range(1, ws.max_row + 1)
        if str(ws.cell(r, 15).value or "").startswith("Guidance (As of")
    ]
    assert len(guidance_rows) == 2
    assert guidance_rows[0] == 7
    assert guidance_rows[1] > guidance_rows[0]
    assert guidance_rows[1] - guidance_rows[0] <= 18
    assert all((ws.row_dimensions[rr].height or 0) >= 19.5 for rr in range(7, 45))
    assert all(ws.cell(guidance_rows[1] - 1, cc).value in (None, "") for cc in range(15, 23))
    operating_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 15).value == "Operating Drivers")
    thesis_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 15).value == "Thesis Bridge")
    assert operating_row > guidance_rows[1]
    assert thesis_row > operating_row
    pe_row = next(r for r in range(thesis_row, ws.max_row + 1) if ws.cell(r, 15).value == "P/E multiple")
    assert ws.cell(pe_row, 21).value == pytest.approx(13.0)
    assert ws.cell(pe_row, 21).fill.fgColor.rgb in {"00FFF2CC", "FFF2CC"}
    output_header = next(r for r in range(thesis_row, ws.max_row + 1) if ws.cell(r, 15).value == "Output")
    output_labels = [
        ws.cell(r, 15).value
        for r in range(output_header + 1, output_header + 13)
    ]
    assert output_labels == [
        "Thesis Adj EBITDA",
        "Thesis FCF",
        "Thesis EPS",
        "EV @ EV/Adj EBITDA",
        "Equity value @ EV/Adj EBITDA",
        "Equity value @ P/E",
        "Equity value @ FCF yield",
        None,
        "Range summary",
        "Value/share @ P/E",
        "Value/share @ EV/Adj EBITDA",
        "Value/share @ FCF yield",
    ]
    range_row = next(r for r in range(output_header, ws.max_row + 1) if ws.cell(r, 15).value == "Range summary")
    value_share_rows = [
        next(r for r in range(output_header, ws.max_row + 1) if ws.cell(r, 15).value == label)
        for label in ("Value/share @ P/E", "Value/share @ EV/Adj EBITDA", "Value/share @ FCF yield")
    ]
    range_formula = str(ws.cell(range_row, 21).value or "")
    assert all(f"U{rr}" in range_formula for rr in value_share_rows)
    assert "MIN(" in range_formula and "MAX(" in range_formula
    assert "U" + str(output_header + 1) not in range_formula


def test_anf_valuation_fiscal_yoy_and_delta_helpers_use_retail_quarters() -> None:
    quarters = pd.to_datetime(
        ["2025-02-01", "2025-05-03", "2025-08-02", "2025-11-01", "2026-01-31"]
    )
    latest = pd.Timestamp("2026-01-31")
    prior = pd.Timestamp("2025-02-01")

    assert _anf_prior_year_quarter(latest, quarters) == prior

    rev = {prior: 1_584_917_000.0, latest: 1_669_802_000.0}
    shares = {prior: 52_461_000.0, latest: 46_837_000.0}
    eps = {prior: 3.57, latest: 3.68}
    fcf = {prior: 256_757_000.0, latest: 250_580_000.0}

    assert _anf_yoy_map_for_fiscal_periods(rev, quarters)[latest] == pytest.approx(0.0536, abs=0.0002)
    assert _anf_value_delta_map_for_fiscal_periods(shares, quarters, comparison="yoy")[latest] == pytest.approx(-5_624_000.0)
    assert _anf_value_delta_map_for_fiscal_periods(eps, quarters, comparison="yoy")[latest] == pytest.approx(0.11)
    assert _anf_value_delta_map_for_fiscal_periods(fcf, quarters, comparison="yoy")[latest] == pytest.approx(-6_177_000.0)


def test_anf_ytd_buybacks_are_normalized_before_ttm_valuation_use() -> None:
    quarters = pd.to_datetime(["2025-05-03", "2025-08-02", "2025-11-01", "2026-01-31"])
    cumulative = {
        pd.Timestamp("2025-05-03"): 200_000_000.0,
        pd.Timestamp("2025-08-02"): 250_000_000.0,
        pd.Timestamp("2025-11-01"): 350_000_000.0,
        pd.Timestamp("2026-01-31"): 450_000_000.0,
    }

    out = _anf_normalize_ytd_buyback_cash_map_for_valuation(cumulative, quarters)

    assert out[pd.Timestamp("2025-05-03")] == pytest.approx(200_000_000.0)
    assert out[pd.Timestamp("2025-08-02")] == pytest.approx(50_000_000.0)
    assert out[pd.Timestamp("2025-11-01")] == pytest.approx(100_000_000.0)
    assert out[pd.Timestamp("2026-01-31")] == pytest.approx(100_000_000.0)
    assert sum(v for v in out.values() if v is not None) == pytest.approx(450_000_000.0)


def test_anf_annual_buyback_summary_is_not_labeled_latest_quarter() -> None:
    q4 = pd.Timestamp("2026-01-31")

    assert _anf_buyback_execution_is_year_or_ttm(
        q4,
        "During the fiscal year, the Company repurchased 5.4 million shares for $450 million.",
        cash_amount=450_000_000.0,
        shares_amount=5_400_000.0,
    )
    summary = _anf_format_year_ttm_buyback_summary(
        q4,
        shares_amount=5_400_000.0,
        cash_amount=450_000_000.0,
        avg_price=83.3333,
    )

    assert summary == "2025 year / TTM buybacks: 5.4m shares for ~$450m at ~$83.33/share"
    assert "Latest quarter" not in summary


def test_anf_net_cash_yoy_flag_uses_net_cash_wording() -> None:
    label, status = _net_debt_yoy_flag_label_and_status_for_position(
        13_200_000.0,
        -759_540_000.0,
    )

    assert label == "Watch: Net cash decreased YoY"
    assert status == "WARN"


def test_anf_valuation_guidance_rows_cover_current_outlook_without_fy_labels() -> None:
    guidance = pd.DataFrame(
        [
            {"quarter": "2026-01-31", "period_label": "FY2026", "period_type": "annual", "metric_hint": "Revenue", "low": 3, "high": 5, "unit": "%", "line": "2026 year net sales growth 3% to 5%"},
            {"quarter": "2026-01-31", "period_label": "FY2026", "period_type": "annual", "metric_hint": "Operating margin", "low": 12.0, "high": 12.5, "unit": "%", "line": "2026 year operating margin 12.0% to 12.5%"},
            {"quarter": "2026-01-31", "period_label": "FY2026", "period_type": "annual", "metric_hint": "Adj EPS", "low": 10.2, "high": 11.0, "unit": "$/share", "line": "2026 year EPS $10.20 to $11.00"},
            {"quarter": "2026-01-31", "period_label": "FY2026", "period_type": "annual", "metric_hint": "Share repurchases", "value": 450, "unit": "$m", "line": "2026 share repurchases around $450m"},
            {"quarter": "2026-01-31", "period_label": "FY2026", "period_type": "annual", "metric_hint": "Diluted shares", "value": 45, "unit": "m shares", "line": "2026 diluted shares around 45m"},
            {"quarter": "2026-01-31", "period_label": "FY2026", "period_type": "annual", "metric_hint": "Capex", "low": 200, "high": 225, "unit": "$m", "line": "2026 capex $200m to $225m"},
            {"quarter": "2026-01-31", "period_label": "Q1 FY2026", "period_type": "quarter", "metric_hint": "Revenue", "low": 1, "high": 3, "unit": "%", "line": "Q1 2026 net sales growth 1% to 3%"},
            {"quarter": "2026-01-31", "period_label": "Q1 FY2026", "period_type": "quarter", "metric_hint": "Operating margin", "value": 7, "unit": "%", "line": "Q1 2026 operating margin around 7%"},
            {"quarter": "2026-01-31", "period_label": "Q1 FY2026", "period_type": "quarter", "metric_hint": "Adj EPS", "low": 1.2, "high": 1.3, "unit": "$/share", "line": "Q1 2026 EPS $1.20 to $1.30"},
            {"quarter": "2026-01-31", "period_label": "Q1 FY2026", "period_type": "quarter", "metric_hint": "Tariffs", "value": 290, "unit": "bps", "line": "Q1 tariff headwind about 290 bps"},
        ]
    )

    rows = _anf_valuation_guidance_rows(guidance)
    row_key = {(r["horizon"], r["metric"]): r["guidance"] for r in rows}

    assert row_key[("2026 year", "Revenue growth")] == "+3-5%"
    assert row_key[("2026 year", "Operating margin")] == "12.0-12.5%"
    assert row_key[("2026 year", "Adj EPS")] == "$10.20-$11.00"
    assert row_key[("2026-Q1", "Revenue growth")] == "+1-3%"
    assert row_key[("2026-Q1", "Tariff headwind")] == "~290 bps"
    assert all("FY" not in str(v) for row in rows for v in row.values())


def test_anf_qa_status_normalizer_fills_blanks_and_downgrades_expected_gaps() -> None:
    checks = pd.DataFrame(
        [
            {"check": "Adj EBITDA (Q)", "status": "", "severity": "", "message": "PASS: source supports latest quarter."},
            {"check": "Debt_Recon", "status": "FAIL", "severity": "FAIL", "message": "Latest-quarter debt coverage total $0.1m vs History_Q debt_core $0.0m."},
            {"check": "hidden_flag_A", "status": "FAIL", "severity": "FAIL", "message": "missing shares_out, shares_yoy"},
            {"check": "hidden_flag_E", "status": "FAIL", "severity": "FAIL", "message": "missing market price / fcf_yield"},
            {"check": "source_coverage", "status": float("nan"), "severity": float("nan"), "message": ""},
        ]
    )

    out = _anf_normalize_qa_status_rows(checks, is_anf_profile=True)

    assert out.loc[0, "status"] == "pass"
    assert out.loc[1, "status"] == "warn"
    assert out.loc[2, "status"] == "warn"
    assert out.loc[3, "status"] == "warn"
    assert out.loc[4, "status"] == "info"
    assert not out["status"].astype(str).str.lower().isin({"", "nan", "none"}).any()
