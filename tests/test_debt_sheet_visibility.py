from __future__ import annotations

from copy import deepcopy
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import Workbook

import pbi_xbrl.excel_writer_financials as financials
from pbi_xbrl.debt_sheet_visibility import (
    DEBT_CREDIT_NOTES_SHEET,
    DEBT_MATURITY_SHEET,
    DEBT_PROFILE_SHEET,
    DebtSheetVisibilityError,
    LEVERAGE_LIQUIDITY_SHEET,
    REVOLVER_HISTORY_SHEET,
    apply_legacy_debt_sheet_visibility,
    debt_profile_source_backed_row_count,
    debt_sheet_default_states,
    debt_sheet_minimum_counts,
    debt_sheet_readiness_contracts,
    mark_debt_maturity_reconciliation,
    mark_debt_profile_readiness,
    resolve_legacy_debt_sheet_visibility,
)
from pbi_xbrl.new_ticker_debt_scope import (
    DEBT_PROFILE_ECONOMIC_VALIDATION_CONTRACT,
    validate_debt_profile_economic_subject,
)
from pbi_xbrl.workbook_modules import load_workbook_module_manifest


CONDITIONAL_SHEETS = {
    DEBT_PROFILE_SHEET,
    REVOLVER_HISTORY_SHEET,
    LEVERAGE_LIQUIDITY_SHEET,
    DEBT_CREDIT_NOTES_SHEET,
    DEBT_MATURITY_SHEET,
}


def _debt_sheet_names() -> tuple[str, ...]:
    payload = load_workbook_module_manifest()
    module = next(row for row in payload["modules"] if row["module_id"] == "debt_liquidity")
    return tuple(str(row["sheet"]) for row in module["sheets"])


def _workbook_with_debt_sheets() -> Workbook:
    workbook = Workbook()
    workbook.active.title = "SUMMARY"
    for name in _debt_sheet_names():
        workbook.create_sheet(name)
    return workbook


def _profile_economic_validation(
    *,
    subject_id: str = "fixture_revolver",
    as_of_date: str = "2026-06-30",
    with_lineage: bool = True,
):
    return validate_debt_profile_economic_subject(
        subject_kind="facility",
        subject_ids=(subject_id,),
        as_of_date=as_of_date,
        evidence_keys=("fixture_facility_evidence",),
        evidence_refs=("fixture.htm#facility",) if with_lineage else (),
        source_refs=("fixture.htm",) if with_lineage else (),
        source_row_refs=("table_1_row_1",) if with_lineage else (),
        source_contract="resolved_debt_facility_disposition",
        economic_validated=True,
    )


def _valid_profile_frame(rows: int = 4) -> pd.DataFrame:
    metrics = ("debt_core", "cash", "lease_liabilities", "debt_fair_value")
    frame = pd.DataFrame(
        [
            {
                "quarter": "2026-06-30",
                "metric": metrics[index] if index < len(metrics) else f"source_metric_{index}",
                "value": float(index + 1),
                "source": "History_Q",
                "note": "source-backed",
            }
            for index in range(rows)
        ]
    )
    return mark_debt_profile_readiness(
        frame,
        economic_validation=_profile_economic_validation(),
    )


def _valid_revolver_frame(periods: int = 4) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "quarter": f"202{index + 1}-12-31",
                "revolver_commitment": 100_000_000.0 + index,
                "commitment_source_type": "xbrl",
            }
            for index in range(periods)
        ]
    )


def _valid_leverage_frame(periods: int = 4) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "quarter": f"202{index + 1}-12-31",
                "corporate_net_debt": 50_000_000.0 + index,
                "ebitda_ttm": 25_000_000.0,
                "corporate_net_leverage": (50_000_000.0 + index) / 25_000_000.0,
                "corporate_net_leverage_basis": "gaap_ebitda_ttm",
            }
            for index in range(periods)
        ]
    )


def _valid_credit_notes_frame(rows: int = 2) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "quarter": "2026-06-30",
                "category": f"note_type_{index}",
                "snippet": "Exact bounded source note.",
                "source_class": "filing_text",
                "method": "debt_text_scan",
                "qa_severity": "pass",
            }
            for index in range(rows)
        ]
    )


def _valid_maturity_frame() -> pd.DataFrame:
    frame = pd.DataFrame(
        [
            {
                "quarter": "2026-06-30",
                "maturity_year": 2028,
                "maturity_label": "2028",
                "amount_total": 125_000_000.0,
                "source_kind": "Debt_Tranches_Latest",
                "source_basis": "principal_tranche_sum",
            },
            {
                "quarter": "2026-06-30",
                "maturity_year": 2030,
                "maturity_label": "2030",
                "amount_total": 75_000_000.0,
                "source_kind": "Debt_Tranches_Latest",
                "source_basis": "principal_tranche_sum",
            },
        ]
    )
    return mark_debt_maturity_reconciliation(frame, reconciled=True)


def _ready_sheet_frames() -> dict[str, pd.DataFrame]:
    return {
        "Debt_Tranches_Latest": pd.DataFrame([{"private": "support"}]),
        REVOLVER_HISTORY_SHEET: _valid_revolver_frame(),
        DEBT_PROFILE_SHEET: _valid_profile_frame(),
        DEBT_MATURITY_SHEET: _valid_maturity_frame(),
        "Debt_Buckets": pd.DataFrame([{"private": "support"}]),
        "Debt_Recon": pd.DataFrame([{"private": "support"}]),
        "Debt_Tranches_Q": pd.DataFrame([{"private": "support"}]),
        DEBT_CREDIT_NOTES_SHEET: _valid_credit_notes_frame(),
        LEVERAGE_LIQUIDITY_SHEET: _valid_leverage_frame(),
    }


def test_manifest_is_single_owner_for_debt_readiness_thresholds() -> None:
    minimums = debt_sheet_minimum_counts(
        (
            DEBT_PROFILE_SHEET,
            REVOLVER_HISTORY_SHEET,
            LEVERAGE_LIQUIDITY_SHEET,
            DEBT_CREDIT_NOTES_SHEET,
        )
    )
    assert minimums == {
        DEBT_PROFILE_SHEET: 4,
        REVOLVER_HISTORY_SHEET: 4,
        LEVERAGE_LIQUIDITY_SHEET: 4,
        DEBT_CREDIT_NOTES_SHEET: 2,
    }


def test_manifest_explicitly_owns_hidden_baseline_and_ready_visibility_overlay() -> None:
    defaults = debt_sheet_default_states()
    contracts = debt_sheet_readiness_contracts(tuple(sorted(CONDITIONAL_SHEETS)))

    assert all(defaults[sheet_name] == "hidden" for sheet_name in contracts)
    assert all(contract.visibility_mode == "readiness_overlay" for contract in contracts.values())
    assert all(contract.ready_state == "visible" for contract in contracts.values())
    assert (
        contracts[DEBT_PROFILE_SHEET].economic_validation_contract
        == DEBT_PROFILE_ECONOMIC_VALIDATION_CONTRACT
    )
    assert all(
        not contract.economic_validation_contract
        for sheet_name, contract in contracts.items()
        if sheet_name != DEBT_PROFILE_SHEET
    )


def test_resolver_fails_closed_when_manifest_ready_state_conflicts_with_overlay_contract() -> None:
    payload = deepcopy(load_workbook_module_manifest())
    module = next(row for row in payload["modules"] if row["module_id"] == "debt_liquidity")
    profile = next(row for row in module["sheets"] if row["sheet"] == DEBT_PROFILE_SHEET)
    profile["readiness"]["ready_state"] = "hidden"

    with pytest.raises(DebtSheetVisibilityError, match="unsupported ready state"):
        resolve_legacy_debt_sheet_visibility(_ready_sheet_frames(), module_payload=payload)


def test_profile_requires_independent_economics_before_presentation_rows_can_publish() -> None:
    frame = mark_debt_profile_readiness(
        pd.DataFrame(
            [
                {
                    "quarter": "2026-06-30",
                    "metric": f"derived_metric_{index}",
                    "value": float(index + 1),
                    "source": "Derived",
                }
                for index in range(4)
            ]
        ),
        economic_validation=_profile_economic_validation(),
    )

    assert debt_profile_source_backed_row_count(frame) == 0
    assert resolve_legacy_debt_sheet_visibility(
        {**_ready_sheet_frames(), DEBT_PROFILE_SHEET: frame}
    )[DEBT_PROFILE_SHEET] == "hidden"


@pytest.mark.parametrize(
    "source",
    ("History_Q", "SEC filing"),
    ids=("history_q", "otherwise_source_backed"),
)
def test_four_arbitrary_metrics_without_validated_facility_fail_closed(source: str) -> None:
    frame = pd.DataFrame(
        [
            {
                "quarter": "2026-06-30",
                "metric": f"unrelated_metric_{index}",
                "value": float(index + 1),
                "source": source,
            }
            for index in range(4)
        ]
    )

    assert debt_profile_source_backed_row_count(frame) == 4
    assert resolve_legacy_debt_sheet_visibility(
        {**_ready_sheet_frames(), DEBT_PROFILE_SHEET: frame}
    )[DEBT_PROFILE_SHEET] == "hidden"


def test_profile_validation_without_source_occurrence_fails_closed() -> None:
    frame = _valid_profile_frame()
    invalid = _profile_economic_validation(with_lineage=False)
    assert not invalid.passed
    mark_debt_profile_readiness(frame, economic_validation=invalid)

    assert resolve_legacy_debt_sheet_visibility(
        {**_ready_sheet_frames(), DEBT_PROFILE_SHEET: frame}
    )[DEBT_PROFILE_SHEET] == "hidden"


def test_profile_validation_date_must_match_latest_presentation_period() -> None:
    frame = _valid_profile_frame()
    mark_debt_profile_readiness(
        frame,
        economic_validation=_profile_economic_validation(as_of_date="2026-03-31"),
    )

    assert resolve_legacy_debt_sheet_visibility(
        {**_ready_sheet_frames(), DEBT_PROFILE_SHEET: frame}
    )[DEBT_PROFILE_SHEET] == "hidden"


def test_conflicting_profile_and_revolver_economic_identities_fail_closed() -> None:
    frames = _ready_sheet_frames()
    frames[REVOLVER_HISTORY_SHEET].attrs.update(frames[DEBT_PROFILE_SHEET].attrs)
    mark_debt_profile_readiness(
        frames[REVOLVER_HISTORY_SHEET],
        economic_validation=_profile_economic_validation(subject_id="different_revolver"),
    )

    with pytest.raises(DebtSheetVisibilityError, match="conflicting economic validation"):
        resolve_legacy_debt_sheet_visibility(frames)


def test_profile_visibility_is_source_order_independent() -> None:
    frames = _ready_sheet_frames()
    forward = resolve_legacy_debt_sheet_visibility(frames)
    reversed_profile = frames[DEBT_PROFILE_SHEET].iloc[::-1].reset_index(drop=True)
    reversed_profile.attrs.update(frames[DEBT_PROFILE_SHEET].attrs)
    reverse = resolve_legacy_debt_sheet_visibility(
        {**frames, DEBT_PROFILE_SHEET: reversed_profile}
    )

    assert forward == reverse


def test_revolver_rejects_unrelated_source_column_without_value_source_pair() -> None:
    frame = pd.DataFrame(
        [
            {
                "quarter": f"202{index + 1}-12-31",
                "revolver_commitment": 100_000_000.0 + index,
                "commitment_source_type": "missing",
                "availability_source_type": "xbrl",
            }
            for index in range(4)
        ]
    )

    assert resolve_legacy_debt_sheet_visibility(
        {**_ready_sheet_frames(), REVOLVER_HISTORY_SHEET: frame}
    )[REVOLVER_HISTORY_SHEET] == "hidden"


def test_leverage_rejects_four_cash_only_rows_without_complete_disposition() -> None:
    frame = pd.DataFrame(
        [
            {
                "quarter": f"202{index + 1}-12-31",
                "cash": 50_000_000.0 + index,
                "liquidity": 50_000_000.0 + index,
            }
            for index in range(4)
        ]
    )

    assert resolve_legacy_debt_sheet_visibility(
        {**_ready_sheet_frames(), LEVERAGE_LIQUIDITY_SHEET: frame}
    )[LEVERAGE_LIQUIDITY_SHEET] == "hidden"


def test_all_debt_sheets_follow_manifest_and_non_debt_sheet_is_untouched() -> None:
    workbook = _workbook_with_debt_sheets()
    try:
        states = apply_legacy_debt_sheet_visibility(workbook, _ready_sheet_frames())

        assert set(states) == set(_debt_sheet_names())
        assert all(workbook[name].sheet_state == "visible" for name in CONDITIONAL_SHEETS)
        assert all(
            workbook[name].sheet_state == "hidden"
            for name in set(_debt_sheet_names()) - CONDITIONAL_SHEETS
        )
        assert workbook["SUMMARY"].sheet_state == "visible"
    finally:
        workbook.close()


def test_anf_two_row_profile_hides_only_profile_among_ready_products() -> None:
    frames = _ready_sheet_frames()
    frames[DEBT_PROFILE_SHEET] = _valid_profile_frame(rows=2)
    states = resolve_legacy_debt_sheet_visibility(frames)

    assert states[DEBT_PROFILE_SHEET] == "hidden"
    assert all(states[name] == "visible" for name in CONDITIONAL_SHEETS - {DEBT_PROFILE_SHEET})


@pytest.mark.parametrize(
    ("sheet_name", "insufficient"),
    (
        (DEBT_PROFILE_SHEET, _valid_profile_frame(rows=3)),
        (REVOLVER_HISTORY_SHEET, _valid_revolver_frame(periods=3)),
        (LEVERAGE_LIQUIDITY_SHEET, _valid_leverage_frame(periods=3)),
        (DEBT_CREDIT_NOTES_SHEET, _valid_credit_notes_frame(rows=1)),
    ),
)
def test_each_conditional_threshold_fails_closed_without_suppressing_other_products(
    sheet_name: str,
    insufficient: pd.DataFrame,
) -> None:
    frames = _ready_sheet_frames()
    frames[sheet_name] = insufficient
    states = resolve_legacy_debt_sheet_visibility(frames)

    assert states[sheet_name] == "hidden"
    assert all(states[name] == "visible" for name in CONDITIONAL_SHEETS - {sheet_name})


@pytest.mark.parametrize(
    "frame",
    [
        pd.DataFrame(),
        pd.DataFrame([{"message": "No data for current build"}]),
        mark_debt_maturity_reconciliation(
            pd.DataFrame(
                [
                    {
                        "quarter": "2026-06-30",
                        "maturity_year": None,
                        "maturity_label": "Needs review: tranche tie-out failed",
                        "amount_total": None,
                        "source_kind": "qa_guardrail",
                        "source_basis": "",
                    }
                ]
            ),
            reconciled=False,
        ),
    ],
    ids=("empty", "placeholder", "qa_guardrail"),
)
def test_non_publishable_maturity_hides_only_maturity(frame: pd.DataFrame) -> None:
    frames = _ready_sheet_frames()
    frames[DEBT_MATURITY_SHEET] = frame
    states = resolve_legacy_debt_sheet_visibility(frames)

    assert states[DEBT_MATURITY_SHEET] == "hidden"
    assert all(states[name] == "visible" for name in CONDITIONAL_SHEETS - {DEBT_MATURITY_SHEET})


def test_incomplete_maturity_stays_hidden_even_with_reconciled_marker() -> None:
    frames = _ready_sheet_frames()
    frame = _valid_maturity_frame()
    frame.loc[0, "amount_total"] = None
    frames[DEBT_MATURITY_SHEET] = frame

    assert resolve_legacy_debt_sheet_visibility(frames)[DEBT_MATURITY_SHEET] == "hidden"


def test_missing_debt_sheets_are_ignored_when_applying_and_unrelated_sheet_is_untouched() -> None:
    workbook = Workbook()
    workbook.active.title = "Unrelated"
    try:
        states = apply_legacy_debt_sheet_visibility(workbook, _ready_sheet_frames())

        assert set(states) == set(_debt_sheet_names())
        assert workbook.sheetnames == ["Unrelated"]
        assert workbook["Unrelated"].sheet_state == "visible"
    finally:
        workbook.close()


def test_legacy_debt_writer_consumes_complete_manifest_visibility_owner(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = Workbook()
    workbook.active.title = "SUMMARY"
    frames = _ready_sheet_frames()

    def write_sheet(name: str, frame: pd.DataFrame) -> None:
        worksheet = workbook.create_sheet(name)
        if frame.empty:
            worksheet["A1"] = "No data for current build"
            return
        worksheet.append(list(frame.columns))
        for row in frame.itertuples(index=False, name=None):
            worksheet.append(list(row))

    inputs = SimpleNamespace(
        debt_tranches_latest=frames["Debt_Tranches_Latest"],
        revolver_history=frames[REVOLVER_HISTORY_SHEET],
        debt_profile=frames[DEBT_PROFILE_SHEET],
        debt_maturity=frames[DEBT_MATURITY_SHEET],
        debt_buckets=frames["Debt_Buckets"],
        debt_recon=frames["Debt_Recon"],
        debt_tranches=frames["Debt_Tranches_Q"],
        debt_credit_notes=frames[DEBT_CREDIT_NOTES_SHEET],
    )
    context = SimpleNamespace(
        wb=workbook,
        inputs=inputs,
        callbacks=SimpleNamespace(write_sheet=write_sheet),
        require_derived_frame=lambda name: frames[LEVERAGE_LIQUIDITY_SHEET],
    )
    monkeypatch.setattr(financials, "ensure_valuation_inputs", lambda _ctx: None)
    try:
        financials.write_debt_sheets(context)

        assert all(workbook[name].sheet_state == "visible" for name in CONDITIONAL_SHEETS)
        assert all(
            workbook[name].sheet_state == "hidden"
            for name in set(_debt_sheet_names()) - CONDITIONAL_SHEETS
        )
        assert workbook["SUMMARY"].sheet_state == "visible"
    finally:
        workbook.close()
