from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
import pytest

from pbi_xbrl.excel_formula_serialization import (
    FormulaSerializationError,
    inventory_workbook_formulas,
    inventory_xlsx_formula_xml,
    serialize_formula_expression,
    serialize_workbook_formulas_for_ooxml,
    validate_workbook_formula_compatibility,
    validate_xlsx_formula_compatibility,
)
from pbi_xbrl.new_ticker_binding_planner import reproduce_binding_plan
from pbi_xbrl.new_ticker_style_planner import StylePlan
from pbi_xbrl.new_ticker_value_filler import fill_standard_template_from_package


ROOT = Path(__file__).resolve().parents[1]
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
PACKAGE = ROOT / "tests" / "fixtures" / "normalized_packages" / "TEST_minimal_valid.json"
SHELL_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"


def test_token_aware_future_function_serialization_is_bounded_and_idempotent() -> None:
    logical = (
        '=LET(routeValue,MAXIFS(A1:A4,B1:B4,"MAXIFS(A1)"),'
        'routeValue+MINIFS(C1:C4,D1:D4,1)+SUMIFS(E1:E4,F1:F4,1))'
    )
    expected = (
        '=_xlfn.LET(_xlpm.routeValue,_xlfn.MAXIFS(A1:A4,B1:B4,"MAXIFS(A1)"),'
        '_xlpm.routeValue+_xlfn.MINIFS(C1:C4,D1:D4,1)+SUMIFS(E1:E4,F1:F4,1))'
    )

    serialized = serialize_formula_expression(logical)

    assert serialized == expected
    assert serialize_formula_expression(serialized) == serialized
    assert serialize_formula_expression("='MAXIFS data'!A1+Table1[MAXIFS]+MAXIFS_Name") == "='MAXIFS data'!A1+Table1[MAXIFS]+MAXIFS_Name"
    assert serialize_formula_expression('=IF(A1="LET(x,1,x)",SUMIFS(B:B,C:C,1),0)') == '=IF(A1="LET(x,1,x)",SUMIFS(B:B,C:C,1),0)'


@pytest.mark.parametrize(
    "formula",
    (
        "=XLOOKUP(A1,B:B,C:C)",
        "=_xlfn.SUMIFS(A:A,B:B,1)",
        "=_xlfn.LET(_xlpm.x,1,_xlpm.y)",
        "=LET(x,1)",
    ),
)
def test_unknown_functions_bad_prefixes_and_malformed_let_fail_closed(formula: str) -> None:
    with pytest.raises(FormulaSerializationError):
        serialize_formula_expression(formula)


def test_workbook_serializer_covers_cells_names_validations_and_conditional_formulas(tmp_path: Path) -> None:
    workbook = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        ws = workbook["SUMMARY"]
        ws["Z1"] = "=MAXIFS(A:A,B:B,1)"
        validation = DataValidation(type="custom", formula1="=LET(x,1,x=1)")
        validation.add(ws["Z2"])
        ws.add_data_validation(validation)
        ws.conditional_formatting.add(
            "Z3",
            FormulaRule(formula=["MINIFS(A:A,B:B,1)>0"], fill=PatternFill(fill_type="solid", fgColor="FF00FF00")),
        )
        workbook.defined_names.add(DefinedName("SerializerTest", attr_text="=LET(x,1,x)"))

        serialize_workbook_formulas_for_ooxml(workbook)
        first = inventory_workbook_formulas(workbook)
        serialize_workbook_formulas_for_ooxml(workbook)
        second = inventory_workbook_formulas(workbook)
        output = tmp_path / "serialized-surfaces.xlsx"
        workbook.save(output)
    finally:
        workbook.close()

    assert first == second
    assert first["unprefixed_future_functions"] == {}
    assert first["unsupported_functions"] == {}
    assert first["malformed_expressions"] == []
    assert validate_xlsx_formula_compatibility(output) == []


def test_checked_in_shell_has_exact_future_function_xml_inventory() -> None:
    inventory = inventory_xlsx_formula_xml(SHELL)
    future_cells = set(inventory["future_function_cells"])
    expected_valuation = {
        f"Valuation!{column}{row}"
        for row in (10, 15, 21, 25, 34, 39, 46, 49, 50, 63, 64, 65, 66, 67, 88, 89, 109, 111, 271)
        for column in "BCDEFGHIJKLM"
    }

    assert inventory["cell_formula_count"] == 2133
    assert inventory["function_counts"]["MAXIFS"] == 324
    assert inventory["function_counts"]["MINIFS"] == 324
    assert inventory["function_counts"]["LET"] == 4
    assert inventory["let_local_occurrences"] == 204
    assert inventory["future_function_cell_count"] == 232
    assert future_cells & expected_valuation == expected_valuation
    assert len(expected_valuation) == 228
    assert set(inventory["let_cells"]) == {
        "Valuation_Summary!H2",
        "Valuation_Summary!I2",
        "Valuation_Summary!J2",
        "Valuation_Summary!K2",
    }
    assert inventory["unprefixed_future_functions"] == {}
    assert inventory["unsupported_functions"] == {}
    assert inventory["malformed_expressions"] == []


def test_exact_cell_filler_preserves_serialized_formula_xml(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    import json

    manifest = json.loads(SHELL_MANIFEST.read_text(encoding="utf-8"))
    bindings = json.loads(BINDING_MAP.read_text(encoding="utf-8"))
    package = json.loads(PACKAGE.read_text(encoding="utf-8"))
    value_plan = reproduce_binding_plan(
        package,
        manifest=manifest,
        binding_payload=bindings,
        shell_path=SHELL,
    )
    no_style_actions = StylePlan(
        ticker=value_plan.ticker,
        module_profile_id="full_union",
        style_contract_digest="0" * 64,
        binding_plan_digest="0" * 64,
    )
    monkeypatch.setattr(
        "pbi_xbrl.new_ticker_value_filler.reproduce_style_plan",
        lambda *_args, **_kwargs: (value_plan, no_style_actions),
    )
    output = tmp_path / "filled.xlsx"

    fill_standard_template_from_package(PACKAGE, output_path=output)

    assert output.exists()
    assert validate_xlsx_formula_compatibility(output) == []
    inventory = inventory_xlsx_formula_xml(output)
    assert inventory["function_counts"]["MAXIFS"] == 324
    assert inventory["function_counts"]["MINIFS"] == 324
    assert inventory["function_counts"]["LET"] == 4
    workbook = load_workbook(output, data_only=False, read_only=False)
    try:
        assert validate_workbook_formula_compatibility(workbook) == []
        assert all(ws.protection.sheet for ws in workbook.worksheets)
        assert sum(
            cell.protection.locked is False
            for ws in workbook.worksheets
            for cell in ws._cells.values()
        ) == 122
    finally:
        workbook.close()
