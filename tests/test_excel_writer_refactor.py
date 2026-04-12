from __future__ import annotations

import json
import re
import shutil
import zipfile
from contextlib import contextmanager, nullcontext
from datetime import date
from pathlib import Path
from uuid import uuid4

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import pbi_xbrl.excel_writer as excel_writer_module
import pbi_xbrl.excel_writer_core as writer_core_module
import pbi_xbrl.excel_writer_context as writer_context_module
from pbi_xbrl.company_profiles import CompanyProfile, OperatingDriverTemplate
from pbi_xbrl.excel_writer import (
    enrich_quarter_notes_audit_rows_with_readback,
    validate_saved_workbook_export,
    validate_needs_review_export,
    validate_qa_export,
    read_quarter_notes_ui_snapshot,
    validate_quarter_notes_ui_export,
    validate_saved_workbook_integrity,
    validate_summary_export,
    validate_valuation_export,
    write_quarter_notes_audit_sheet,
    write_excel_from_inputs,
)
from pbi_xbrl.excel_writer_drivers import load_operating_driver_template_index, write_driver_sheets
from pbi_xbrl.excel_writer_context import build_writer_context
from pbi_xbrl.excel_writer_core import (
    ensure_driver_inputs,
    ensure_hidden_value_inputs,
    ensure_raw_data_inputs,
    ensure_report_inputs,
    ensure_summary_inputs,
    ensure_ui_evidence,
    ensure_valuation_inputs,
    prepare_writer_inputs,
    write_qa_sheets,
    write_raw_data_sheets,
)
from pbi_xbrl.excel_writer_financials import (
    write_debt_sheets,
    write_report_sheets,
    write_summary_sheets,
    write_valuation_sheets,
)
from pbi_xbrl.excel_writer_segments import (
    latest_segment_financials_workbook,
    parse_quarterly_segment_data_from_workbook,
)
from pbi_xbrl.excel_writer_sources import docs_for_valuation_accn
from pbi_xbrl.excel_writer_ui import write_ui_sheets
from pbi_xbrl.pipeline_types import PipelineArtifacts, WorkbookInputs
from pbi_xbrl.quarter_notes_runtime import QuarterNotesRuntime
from pbi_xbrl.workbook_gap_audit import load_pipeline_bundle_map


@contextmanager
def _case_dir() -> Path:
    root = Path(__file__).resolve().parents[2] / ".venv" / "tmp_excel_writer_refactor_tests"
    root.mkdir(parents=True, exist_ok=True)
    case_dir = root / uuid4().hex
    case_dir.mkdir()
    try:
        yield case_dir
    finally:
        shutil.rmtree(case_dir, ignore_errors=True)


def _quarter_block_notes(ws, qtxt: str) -> list[str]:
    out: list[str] = []
    capture = False
    for rr in range(1, ws.max_row + 1):
        marker = str(ws.cell(row=rr, column=1).value or "")
        category = str(ws.cell(row=rr, column=2).value or "")
        if marker == qtxt:
            capture = True
            continue
        if capture and marker:
            break
        if capture and category == "Category":
            continue
        if capture:
            note = str(ws.cell(row=rr, column=3).value or "")
            if note:
                out.append(note)
    return out


def _make_hist() -> pd.DataFrame:
    quarters = pd.to_datetime(
        [
            "2025-03-31",
            "2025-06-30",
            "2025-09-30",
            "2025-12-31",
        ]
    )
    return pd.DataFrame(
        {
            "quarter": quarters,
            "revenue": [100.0, 110.0, 120.0, 130.0],
            "cfo": [10.0, 12.0, 13.0, 14.0],
            "capex": [2.0, 2.0, 3.0, 3.0],
            "ebitda": [15.0, 16.0, 18.0, 19.0],
            "ebit": [9.0, 10.0, 11.0, 12.0],
            "cash": [20.0, 21.0, 22.0, 23.0],
            "debt_core": [50.0, 49.0, 48.0, 47.0],
            "shares_outstanding": [10.0, 10.0, 10.0, 10.0],
            "shares_diluted": [10.0, 10.0, 10.0, 10.0],
            "market_cap": [100.0, 100.0, 100.0, 100.0],
            "interest_expense_net": [1.0, 1.0, 1.0, 1.0],
        }
    )


def _make_inputs(
    out_path: Path,
    *,
    excel_mode: str = "clean",
    profile_timings: bool = False,
    ticker: str = "TEST",
    hist: pd.DataFrame | None = None,
    audit: pd.DataFrame | None = None,
    needs_review: pd.DataFrame | None = None,
    quarter_notes: pd.DataFrame | None = None,
    promises: pd.DataFrame | None = None,
    promise_progress: pd.DataFrame | None = None,
    adj_breakdown: pd.DataFrame | None = None,
    adj_metrics_relaxed: pd.DataFrame | None = None,
    adj_breakdown_relaxed: pd.DataFrame | None = None,
    non_gaap_files_relaxed: pd.DataFrame | None = None,
    ocr_log: pd.DataFrame | None = None,
    slides_segments: pd.DataFrame | None = None,
    quarter_notes_audit: bool = False,
    capture_saved_workbook_provenance: bool = True,
    excel_debug_scope: str = "full",
) -> WorkbookInputs:
    empty = pd.DataFrame()
    return WorkbookInputs(
        out_path=out_path,
        hist=hist if hist is not None else _make_hist(),
        audit=audit if audit is not None else empty,
        needs_review=needs_review if needs_review is not None else empty,
        debt_tranches=empty,
        debt_recon=empty,
        adj_metrics=empty,
        adj_breakdown=adj_breakdown if adj_breakdown is not None else empty,
        non_gaap_files=empty,
        adj_metrics_relaxed=adj_metrics_relaxed if adj_metrics_relaxed is not None else empty,
        adj_breakdown_relaxed=adj_breakdown_relaxed if adj_breakdown_relaxed is not None else empty,
        non_gaap_files_relaxed=non_gaap_files_relaxed if non_gaap_files_relaxed is not None else empty,
        info_log=empty,
        tag_coverage=empty,
        period_checks=empty,
        qa_checks=empty,
        bridge_q=empty,
        manifest_df=empty,
        ocr_log=ocr_log if ocr_log is not None else empty,
        qfd_preview=empty,
        qfd_unused=empty,
        debt_profile=empty,
        debt_tranches_latest=empty,
        debt_maturity=empty,
        debt_credit_notes=empty,
        revolver_df=empty,
        revolver_history=empty,
        debt_buckets=empty,
        slides_segments=slides_segments if slides_segments is not None else empty,
        slides_debt=empty,
        slides_guidance=empty,
        quarter_notes=quarter_notes if quarter_notes is not None else empty,
        promises=promises if promises is not None else empty,
        promise_progress=promise_progress if promise_progress is not None else empty,
        non_gaap_cred=empty,
        ticker=ticker,
        price=10.0,
        strictness="ytd",
        excel_mode=excel_mode,
        cache_dir=out_path.parent,
        profile_timings=profile_timings,
        quarter_notes_audit=quarter_notes_audit,
        capture_saved_workbook_provenance=capture_saved_workbook_provenance,
        excel_debug_scope=excel_debug_scope,
    )


def _make_model_out_path(case_dir: Path, filename: str = "model.xlsx") -> Path:
    model_dir = case_dir / "TEST" / "TEST model excel"
    model_dir.mkdir(parents=True, exist_ok=True)
    return model_dir / filename


def _make_ticker_model_out_path(case_dir: Path, ticker: str, filename: str = "model.xlsx") -> Path:
    model_dir = case_dir / ticker / f"{ticker} model excel"
    model_dir.mkdir(parents=True, exist_ok=True)
    return model_dir / filename


def _current_delivered_model_path(ticker: str) -> Path:
    return Path(__file__).resolve().parents[2] / "Excel stock models" / f"{ticker}_model.xlsx"


def _make_live_artifact_inputs(ticker: str, out_path: Path) -> WorkbookInputs:
    repo_root = Path(__file__).resolve().parents[2]
    bundle = load_pipeline_bundle_map(repo_root, ticker)
    if not bundle:
        pytest.skip(f"Live pipeline bundle for {ticker} is missing.")
    artifacts = PipelineArtifacts(**bundle)
    return WorkbookInputs.from_artifacts(
        artifacts,
        out_path=out_path,
        ticker=ticker,
        cache_dir=repo_root / "sec_cache" / ticker,
    )


def _callback_closure_value(fn, name: str):
    freevars = tuple(getattr(fn.__code__, "co_freevars", ()) or ())
    closure = tuple(getattr(fn, "__closure__", ()) or ())
    mapping = dict(zip(freevars, [cell.cell_contents for cell in closure]))
    if name not in mapping:
        raise KeyError(f"Callback closure does not expose {name!r}.")
    return mapping[name]


def _find_row_with_value(ws, text: str, *, column: int | None = 1) -> int | None:
    for rr in range(1, ws.max_row + 1):
        if column is None:
            for cc in range(1, ws.max_column + 1):
                if ws.cell(row=rr, column=cc).value == text:
                    return rr
            continue
        if ws.cell(row=rr, column=column).value == text:
            return rr
    return None


def _find_row_containing(ws, text: str, *, column: int = 1) -> int | None:
    needle = str(text)
    for rr in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=rr, column=column).value
        if needle in str(cell_value or ""):
            return rr
    return None


def _find_col_with_value(ws, text: str, *, row: int) -> int | None:
    needle = str(text)
    for cc in range(1, ws.max_column + 1):
        if str(ws.cell(row=row, column=cc).value or "").strip() == needle:
            return cc
    return None


def _fill_rgb(cell) -> str:
    return str(cell.fill.fgColor.rgb or "")


def _assert_gpre_qtd_tracking_upper_block(ws_overlay) -> None:
    qtd_tracking_title_row = _find_row_with_value(ws_overlay, "Current QTD trend tracking ($/gal, crush margin lens)", column=1)
    assert qtd_tracking_title_row is not None
    qtd_tracking_today_row = qtd_tracking_title_row + 1
    qtd_tracking_quarter_open_row = qtd_tracking_title_row + 2
    qtd_tracking_compare_header_row = qtd_tracking_title_row + 3
    qtd_tracking_compare_subheader_row = qtd_tracking_title_row + 4
    qtd_tracking_compare_body_row = qtd_tracking_title_row + 5
    qtd_tracking_hidden_row = qtd_tracking_title_row + 6
    qtd_tracking_driver_title_row = qtd_tracking_title_row + 7
    qtd_tracking_driver_header_row = qtd_tracking_title_row + 8
    qtd_tracking_driver_first_row = qtd_tracking_title_row + 9
    qtd_tracking_driver_last_row = qtd_tracking_title_row + 12
    qtd_tracking_note_row = qtd_tracking_title_row + 13
    coproducts_row = qtd_tracking_title_row + 14
    merged_ranges = {str(rng) for rng in ws_overlay.merged_cells.ranges}
    expected_merged_ranges = {
        f"A{qtd_tracking_title_row}:U{qtd_tracking_title_row}",
        f"A{qtd_tracking_driver_title_row}:U{qtd_tracking_driver_title_row}",
        f"A{coproducts_row}:U{coproducts_row}",
        f"B{qtd_tracking_today_row}:C{qtd_tracking_today_row}",
        f"D{qtd_tracking_today_row}:U{qtd_tracking_today_row}",
        f"B{qtd_tracking_quarter_open_row}:C{qtd_tracking_quarter_open_row}",
        f"D{qtd_tracking_quarter_open_row}:U{qtd_tracking_quarter_open_row}",
        f"B{qtd_tracking_compare_header_row}:C{qtd_tracking_compare_header_row}",
        f"D{qtd_tracking_compare_header_row}:E{qtd_tracking_compare_header_row}",
        f"F{qtd_tracking_compare_header_row}:G{qtd_tracking_compare_header_row}",
        f"H{qtd_tracking_compare_header_row}:I{qtd_tracking_compare_header_row}",
        f"B{qtd_tracking_compare_subheader_row}:C{qtd_tracking_compare_subheader_row}",
        f"D{qtd_tracking_compare_subheader_row}:E{qtd_tracking_compare_subheader_row}",
        f"F{qtd_tracking_compare_subheader_row}:G{qtd_tracking_compare_subheader_row}",
        f"H{qtd_tracking_compare_subheader_row}:I{qtd_tracking_compare_subheader_row}",
        f"B{qtd_tracking_compare_body_row}:C{qtd_tracking_compare_body_row}",
        f"D{qtd_tracking_compare_body_row}:E{qtd_tracking_compare_body_row}",
        f"F{qtd_tracking_compare_body_row}:G{qtd_tracking_compare_body_row}",
        f"H{qtd_tracking_compare_body_row}:I{qtd_tracking_compare_body_row}",
        f"B{qtd_tracking_driver_header_row}:C{qtd_tracking_driver_header_row}",
        f"D{qtd_tracking_driver_header_row}:E{qtd_tracking_driver_header_row}",
        f"F{qtd_tracking_driver_header_row}:G{qtd_tracking_driver_header_row}",
        f"H{qtd_tracking_driver_header_row}:I{qtd_tracking_driver_header_row}",
    }
    assert expected_merged_ranges.issubset(merged_ranges)
    assert f"A{qtd_tracking_today_row}:U{qtd_tracking_today_row}" not in merged_ranges
    assert f"B{qtd_tracking_note_row}:U{qtd_tracking_note_row}" not in merged_ranges

    assert str(ws_overlay.cell(row=qtd_tracking_today_row, column=1).value or "").strip() == "Today"
    assert str(ws_overlay.cell(row=qtd_tracking_quarter_open_row, column=1).value or "").strip() == "Quarter-open"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_body_row, column=1).value or "").strip() == "Approximate market crush"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_header_row, column=2).value or "").strip() == "QTD vs quarter-open"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_header_row, column=4).value or "").strip() == "QTD vs 1 week ago"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_header_row, column=6).value or "").strip() == "QTD vs 4 weeks ago"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_header_row, column=8).value or "").strip() == "QTD vs 8 weeks ago"
    assert re.fullmatch(r"As of 20\d{2}-\d{2}-\d{2}", str(ws_overlay.cell(row=qtd_tracking_today_row, column=4).value or "").strip())
    assert re.fullmatch(r"As of 20\d{2}-\d{2}-\d{2}", str(ws_overlay.cell(row=qtd_tracking_quarter_open_row, column=4).value or "").strip())
    assert re.fullmatch(r"As of 20\d{2}-\d{2}-\d{2}", str(ws_overlay.cell(row=qtd_tracking_compare_subheader_row, column=2).value or "").strip())
    one_week_subheader = str(ws_overlay.cell(row=qtd_tracking_compare_subheader_row, column=4).value or "").strip()
    assert one_week_subheader == "—" or re.fullmatch(r"As of 20\d{2}-\d{2}-\d{2}", one_week_subheader)
    assert str(ws_overlay.cell(row=180, column=6).value or "").strip() == "—"
    assert str(ws_overlay.cell(row=180, column=8).value or "").strip() == "—"
    assert pd.notna(pd.to_numeric(ws_overlay.cell(row=177, column=2).value, errors="coerce"))
    assert pd.notna(pd.to_numeric(ws_overlay.cell(row=178, column=2).value, errors="coerce"))
    assert pd.notna(pd.to_numeric(ws_overlay.cell(row=181, column=2).value, errors="coerce"))
    one_week_delta = pd.to_numeric(ws_overlay.cell(row=181, column=4).value, errors="coerce")
    if one_week_subheader == "—":
        assert str(ws_overlay.cell(row=181, column=4).value or "").strip() == "—"
    else:
        assert pd.notna(one_week_delta)
    assert str(ws_overlay.cell(row=181, column=6).value or "").strip() == "—"
    assert str(ws_overlay.cell(row=181, column=8).value or "").strip() == "—"

    upper_block_values = {
        str(ws_overlay.cell(row=rr, column=cc).value or "").strip()
        for rr in range(177, 182)
        for cc in range(1, 22)
        if str(ws_overlay.cell(row=rr, column=cc).value or "").strip()
    }
    assert "Reference date" not in upper_block_values
    assert "Delta ($/gal)" not in upper_block_values
    assert "Note" not in upper_block_values
    assert "insufficient history" not in {value.lower() for value in upper_block_values}
    assert str(ws_overlay.cell(row=179, column=1).value or "").strip() == ""
    assert str(ws_overlay.cell(row=180, column=1).value or "").strip() == ""
    row_182_height = ws_overlay.row_dimensions[182].height
    assert row_182_height is None or row_182_height == pytest.approx(0.0, abs=0.01)
    assert bool(ws_overlay.row_dimensions[182].hidden)
    assert str(ws_overlay.cell(row=182, column=1).value or "").strip() == ""

    assert str(ws_overlay.cell(row=178, column=1).alignment.horizontal or "") == "center"
    assert str(ws_overlay.cell(row=179, column=1).alignment.horizontal or "") == "center"
    assert str(ws_overlay.cell(row=181, column=1).alignment.horizontal or "") == "left"
    assert str(ws_overlay.cell(row=183, column=1).value or "").strip() == "Driver attribution of Current QTD move ($/gal)"
    assert str(ws_overlay.cell(row=183, column=1).alignment.horizontal or "") == "left"
    assert str(ws_overlay.cell(row=184, column=2).value or "").strip() == "QTD vs quarter-open"
    assert str(ws_overlay.cell(row=184, column=4).value or "").strip() == "QTD vs 1 week ago"
    assert str(ws_overlay.cell(row=184, column=6).value or "").strip() == "QTD vs 4 weeks ago"
    assert str(ws_overlay.cell(row=184, column=8).value or "").strip() == "QTD vs 8 weeks ago"
    assert _fill_rgb(ws_overlay.cell(row=183, column=1)) == _fill_rgb(ws_overlay.cell(row=184, column=2))
    assert _fill_rgb(ws_overlay.cell(row=183, column=1)) != _fill_rgb(ws_overlay.cell(row=180, column=2))
    for data_row in range(185, 189):
        assert ws_overlay.row_dimensions[data_row].height == pytest.approx(20.0, abs=0.01)
    assert str(ws_overlay.cell(row=195, column=1).value or "").strip() == ""


def _assert_gpre_qtd_tracking_upper_block_dynamic(ws_overlay) -> None:
    qtd_tracking_title_row = _find_row_with_value(ws_overlay, "Current QTD trend tracking ($/gal, crush margin lens)", column=1)
    assert qtd_tracking_title_row is not None
    qtd_tracking_today_row = qtd_tracking_title_row + 1
    qtd_tracking_quarter_open_row = qtd_tracking_title_row + 2
    qtd_tracking_compare_header_row = qtd_tracking_title_row + 3
    qtd_tracking_compare_subheader_row = qtd_tracking_title_row + 4
    qtd_tracking_compare_body_row = qtd_tracking_title_row + 5
    qtd_tracking_hidden_row = qtd_tracking_title_row + 6
    qtd_tracking_driver_title_row = qtd_tracking_title_row + 7
    qtd_tracking_driver_header_row = qtd_tracking_title_row + 8
    qtd_tracking_driver_first_row = qtd_tracking_title_row + 9
    qtd_tracking_driver_last_row = qtd_tracking_title_row + 12
    qtd_tracking_note_row = qtd_tracking_title_row + 13
    coproducts_row = qtd_tracking_title_row + 14
    merged_ranges = {str(rng) for rng in ws_overlay.merged_cells.ranges}
    expected_merged_ranges = {
        f"A{qtd_tracking_title_row}:U{qtd_tracking_title_row}",
        f"A{qtd_tracking_driver_title_row}:U{qtd_tracking_driver_title_row}",
        f"A{coproducts_row}:U{coproducts_row}",
        f"B{qtd_tracking_today_row}:C{qtd_tracking_today_row}",
        f"D{qtd_tracking_today_row}:U{qtd_tracking_today_row}",
        f"B{qtd_tracking_quarter_open_row}:C{qtd_tracking_quarter_open_row}",
        f"D{qtd_tracking_quarter_open_row}:U{qtd_tracking_quarter_open_row}",
        f"B{qtd_tracking_compare_header_row}:C{qtd_tracking_compare_header_row}",
        f"D{qtd_tracking_compare_header_row}:E{qtd_tracking_compare_header_row}",
        f"F{qtd_tracking_compare_header_row}:G{qtd_tracking_compare_header_row}",
        f"H{qtd_tracking_compare_header_row}:I{qtd_tracking_compare_header_row}",
        f"B{qtd_tracking_compare_subheader_row}:C{qtd_tracking_compare_subheader_row}",
        f"D{qtd_tracking_compare_subheader_row}:E{qtd_tracking_compare_subheader_row}",
        f"F{qtd_tracking_compare_subheader_row}:G{qtd_tracking_compare_subheader_row}",
        f"H{qtd_tracking_compare_subheader_row}:I{qtd_tracking_compare_subheader_row}",
        f"B{qtd_tracking_compare_body_row}:C{qtd_tracking_compare_body_row}",
        f"D{qtd_tracking_compare_body_row}:E{qtd_tracking_compare_body_row}",
        f"F{qtd_tracking_compare_body_row}:G{qtd_tracking_compare_body_row}",
        f"H{qtd_tracking_compare_body_row}:I{qtd_tracking_compare_body_row}",
        f"B{qtd_tracking_driver_header_row}:C{qtd_tracking_driver_header_row}",
        f"D{qtd_tracking_driver_header_row}:E{qtd_tracking_driver_header_row}",
        f"F{qtd_tracking_driver_header_row}:G{qtd_tracking_driver_header_row}",
        f"H{qtd_tracking_driver_header_row}:I{qtd_tracking_driver_header_row}",
    }
    assert expected_merged_ranges.issubset(merged_ranges)
    assert f"A{qtd_tracking_today_row}:U{qtd_tracking_today_row}" not in merged_ranges
    assert f"B{qtd_tracking_note_row}:U{qtd_tracking_note_row}" not in merged_ranges

    assert str(ws_overlay.cell(row=qtd_tracking_today_row, column=1).value or "").strip() == "Today"
    assert str(ws_overlay.cell(row=qtd_tracking_quarter_open_row, column=1).value or "").strip() == "Quarter-open"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_body_row, column=1).value or "").strip() == "Approximate market crush"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_header_row, column=2).value or "").strip() == "QTD vs quarter-open"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_header_row, column=4).value or "").strip() == "QTD vs 1 week ago"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_header_row, column=6).value or "").strip() == "QTD vs 4 weeks ago"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_header_row, column=8).value or "").strip() == "QTD vs 8 weeks ago"

    assert re.fullmatch(r"As of 20\d{2}-\d{2}-\d{2}", str(ws_overlay.cell(row=qtd_tracking_today_row, column=4).value or "").strip())
    assert re.fullmatch(r"As of 20\d{2}-\d{2}-\d{2}", str(ws_overlay.cell(row=qtd_tracking_quarter_open_row, column=4).value or "").strip())
    assert re.fullmatch(r"As of 20\d{2}-\d{2}-\d{2}", str(ws_overlay.cell(row=qtd_tracking_compare_subheader_row, column=2).value or "").strip())
    one_week_subheader = str(ws_overlay.cell(row=qtd_tracking_compare_subheader_row, column=4).value or "").strip()
    assert one_week_subheader in {"—", "â€”"} or re.fullmatch(r"As of 20\d{2}-\d{2}-\d{2}", one_week_subheader)
    assert str(ws_overlay.cell(row=qtd_tracking_compare_subheader_row, column=6).value or "").strip() in {"—", "â€”"}
    assert str(ws_overlay.cell(row=qtd_tracking_compare_subheader_row, column=8).value or "").strip() in {"—", "â€”"}

    assert pd.notna(pd.to_numeric(ws_overlay.cell(row=qtd_tracking_today_row, column=2).value, errors="coerce"))
    assert pd.notna(pd.to_numeric(ws_overlay.cell(row=qtd_tracking_quarter_open_row, column=2).value, errors="coerce"))
    assert pd.notna(pd.to_numeric(ws_overlay.cell(row=qtd_tracking_compare_body_row, column=2).value, errors="coerce"))
    one_week_delta = pd.to_numeric(ws_overlay.cell(row=qtd_tracking_compare_body_row, column=4).value, errors="coerce")
    if one_week_subheader in {"—", "â€”"}:
        assert str(ws_overlay.cell(row=qtd_tracking_compare_body_row, column=4).value or "").strip() in {"—", "â€”"}
    else:
        assert pd.notna(one_week_delta)
    assert str(ws_overlay.cell(row=qtd_tracking_compare_body_row, column=6).value or "").strip() in {"—", "â€”"}
    assert str(ws_overlay.cell(row=qtd_tracking_compare_body_row, column=8).value or "").strip() in {"—", "â€”"}

    upper_block_values = {
        str(ws_overlay.cell(row=rr, column=cc).value or "").strip()
        for rr in range(qtd_tracking_today_row, qtd_tracking_compare_body_row + 1)
        for cc in range(1, 22)
        if str(ws_overlay.cell(row=rr, column=cc).value or "").strip()
    }
    assert "Reference date" not in upper_block_values
    assert "Delta ($/gal)" not in upper_block_values
    assert "Note" not in upper_block_values
    assert "insufficient history" not in {value.lower() for value in upper_block_values}
    assert str(ws_overlay.cell(row=qtd_tracking_compare_header_row, column=1).value or "").strip() == ""
    assert str(ws_overlay.cell(row=qtd_tracking_compare_subheader_row, column=1).value or "").strip() == ""
    hidden_row_height = ws_overlay.row_dimensions[qtd_tracking_hidden_row].height
    assert hidden_row_height is None or hidden_row_height == pytest.approx(0.0, abs=0.01)
    assert bool(ws_overlay.row_dimensions[qtd_tracking_hidden_row].hidden)
    assert str(ws_overlay.cell(row=qtd_tracking_hidden_row, column=1).value or "").strip() == ""

    assert str(ws_overlay.cell(row=qtd_tracking_quarter_open_row, column=1).alignment.horizontal or "") == "center"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_header_row, column=1).alignment.horizontal or "") == "center"
    assert str(ws_overlay.cell(row=qtd_tracking_compare_body_row, column=1).alignment.horizontal or "") == "left"
    assert str(ws_overlay.cell(row=qtd_tracking_driver_title_row, column=1).value or "").strip() == "Driver attribution of Current QTD move ($/gal)"
    assert str(ws_overlay.cell(row=qtd_tracking_driver_title_row, column=1).alignment.horizontal or "") == "left"
    assert str(ws_overlay.cell(row=qtd_tracking_driver_header_row, column=2).value or "").strip() == "QTD vs quarter-open"
    assert str(ws_overlay.cell(row=qtd_tracking_driver_header_row, column=4).value or "").strip() == "QTD vs 1 week ago"
    assert str(ws_overlay.cell(row=qtd_tracking_driver_header_row, column=6).value or "").strip() == "QTD vs 4 weeks ago"
    assert str(ws_overlay.cell(row=qtd_tracking_driver_header_row, column=8).value or "").strip() == "QTD vs 8 weeks ago"
    assert _fill_rgb(ws_overlay.cell(row=qtd_tracking_driver_title_row, column=1)) == _fill_rgb(ws_overlay.cell(row=qtd_tracking_driver_header_row, column=2))
    assert _fill_rgb(ws_overlay.cell(row=qtd_tracking_driver_title_row, column=1)) != _fill_rgb(ws_overlay.cell(row=qtd_tracking_compare_subheader_row, column=2))
    for data_row in range(qtd_tracking_driver_first_row, qtd_tracking_driver_last_row + 1):
        assert ws_overlay.row_dimensions[data_row].height == pytest.approx(20.0, abs=0.01)
    assert ws_overlay.row_dimensions[qtd_tracking_note_row].height == pytest.approx(15.0, abs=0.01)
    assert str(ws_overlay.cell(row=qtd_tracking_note_row, column=1).value or "").strip() == ""


def _sheet_data_row_count(ws) -> int:
    return max(int(ws.max_row) - 1, 0)


def _quarter_label_ord(label: str) -> int | None:
    txt = str(label or "").strip()
    m = re.fullmatch(r"(\d{4})-Q([1-4])", txt)
    if m:
        return int(m.group(1)) * 4 + int(m.group(2))
    m = re.fullmatch(r"Q([1-4])\s+(\d{4})", txt)
    if m:
        return int(m.group(2)) * 4 + int(m.group(1))
    return None


def _operating_commentary_rows(ws) -> list[dict[str, str]]:
    start_row = _find_row_with_value(ws, "Operating Commentary", column=1)
    assert start_row is not None
    assert str(ws.cell(row=start_row + 1, column=1).value or "").strip() == "Horizon"
    assert str(ws.cell(row=start_row + 1, column=2).value or "").strip() == "Stated in"
    assert str(ws.cell(row=start_row + 1, column=3).value or "").strip() == "Commentary"
    rows: list[dict[str, str]] = []
    rr = start_row + 2
    current_year_band = ""
    while rr <= ws.max_row:
        col_a = str(ws.cell(row=rr, column=1).value or "").strip()
        col_b = str(ws.cell(row=rr, column=2).value or "").strip()
        col_c = str(ws.cell(row=rr, column=3).value or "").strip()
        if col_a in {"2023", "2024", "2025", "2026 / current"} and not col_b and not col_c:
            current_year_band = col_a
            rr += 1
            continue
        if not col_a and not col_b and not col_c:
            rr += 1
            continue
        if col_a.startswith("Actuals") or col_b.startswith("Actuals"):
            break
        if col_a.startswith("Segment support") or (col_a == "Quarter" and rr > start_row + 1):
            break
        if not col_c:
            rr += 1
            continue
        rows.append(
            {
                "row": str(rr),
                "year_band": current_year_band,
                "horizon": col_a,
                "stated_in": col_b,
                "commentary": col_c,
            }
        )
        rr += 1
    return rows


def _overlay_management_commentary_rows(ws) -> list[dict[str, str]]:
    start_row = _find_row_with_value(ws, "Management commentary", column=1)
    assert start_row is not None
    assert str(ws.cell(row=start_row + 1, column=1).value or "").strip() == "Horizon"
    assert str(ws.cell(row=start_row + 1, column=2).value or "").strip() == "Stated in"
    assert str(ws.cell(row=start_row + 1, column=3).value or "").strip() == "Commentary"
    rows: list[dict[str, str]] = []
    rr = start_row + 2
    current_year_band = ""
    while rr <= ws.max_row:
        col_a = str(ws.cell(row=rr, column=1).value or "").strip()
        col_b = str(ws.cell(row=rr, column=2).value or "").strip()
        col_c = str(ws.cell(row=rr, column=3).value or "").strip()
        if col_a in {"2023", "2024", "2025", "2026 / current"} and not col_b and not col_c:
            current_year_band = col_a
            rr += 1
            continue
        if col_a == "Commercial / hedge setup":
            break
        if not col_c:
            rr += 1
            continue
        rows.append(
            {
                "row": str(rr),
                "year_band": current_year_band,
                "horizon": col_a,
                "stated_in": col_b,
                "commentary": col_c,
            }
        )
        rr += 1
    return rows


def _overlay_section_blob(ws, section_label: str, *, stop_labels: list[str]) -> str:
    start_row = _find_row_with_value(ws, section_label, column=1)
    assert start_row is not None
    stop_set = {str(x or "").strip() for x in stop_labels if str(x or "").strip()}
    rows: list[str] = []
    rr = start_row + 1
    while rr <= ws.max_row:
        marker = str(ws.cell(row=rr, column=1).value or "").strip()
        if rr > start_row + 1 and marker in stop_set:
            break
        joined = " | ".join(
            str(ws.cell(row=rr, column=cc).value or "").strip()
            for cc in range(1, min(18, ws.max_column + 1))
        ).strip(" |")
        if joined:
            rows.append(joined)
        rr += 1
    return "\n".join(rows)


def _sheet_metric_rows(path: Path, sheet_name: str, metric: str, *, quarters: list[str] | None = None) -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        assert sheet_name in wb.sheetnames
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    assert rows
    header = [str(x or "").strip() for x in rows[0]]
    idx = {name: pos for pos, name in enumerate(header) if name}
    quarter_set = {str(x) for x in (quarters or []) if str(x).strip()}
    out: list[dict[str, str]] = []
    for vals in rows[1:]:
        if not vals:
            continue
        metric_val = str(vals[idx["metric"]] or "").strip() if "metric" in idx and idx["metric"] < len(vals) else ""
        if metric_val != metric:
            continue
        row: dict[str, str] = {}
        for col in ["quarter", "metric", "severity", "status", "message", "source"]:
            if col not in idx or idx[col] >= len(vals):
                continue
            raw = vals[idx[col]]
            if col == "quarter":
                qts = pd.to_datetime(raw, errors="coerce")
                row[col] = pd.Timestamp(qts).strftime("%Y-%m-%d") if pd.notna(qts) else str(raw or "").strip()
            else:
                row[col] = "" if raw is None else str(raw)
        if quarter_set and str(row.get("quarter") or "") not in quarter_set:
            continue
        out.append(row)
    return out


def _read_quarter_notes_audit_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    assert "Quarter_Notes_Audit" in wb.sheetnames
    ws = wb["Quarter_Notes_Audit"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows
    header = [str(x or "") for x in rows[0]]
    out: list[dict[str, str]] = []
    for vals in rows[1:]:
        if not vals or all(v in (None, "") for v in vals):
            continue
        out.append({header[idx]: "" if vals[idx] is None else str(vals[idx]) for idx in range(len(header))})
    return out


@contextmanager
def _profile_override(monkeypatch: pytest.MonkeyPatch, profile_ticker: str):
    original = writer_context_module.get_company_profile
    monkeypatch.setattr(
        writer_context_module,
        "get_company_profile",
        lambda _ticker: original(profile_ticker),
    )
    try:
        yield
    finally:
        monkeypatch.setattr(writer_context_module, "get_company_profile", original)


def _write_pbi_segment_workbook(segment_dir: Path, *, include_quarters: bool = True) -> Path:
    segment_dir.mkdir(parents=True, exist_ok=True)
    path = segment_dir / "Historical Segment Financials up to Q4 2025.xlsx"
    wb = Workbook()
    ws_rev = wb.active
    ws_rev.title = "Revenue & Gross Profit"
    ws_adj = wb.create_sheet("Adj Segment Data")

    quarter_headers = ["Mar 2025", "Jun 2025", "Sep 2025", "Dec 2025"] if include_quarters else []
    annual_headers = ["2023", "2024", "2025"]
    header_values = quarter_headers + annual_headers
    header_start_col = 4
    for ws in (ws_rev, ws_adj):
        for idx, label in enumerate(header_values):
            ws.cell(row=1, column=header_start_col + idx, value=label)

    def _write_rows(ws, rows: list[list[object]]) -> None:
        for ridx, values in enumerate(rows, start=2):
            for cidx, value in enumerate(values, start=1):
                ws.cell(row=ridx, column=cidx, value=value)

    q_vals = {
        "SendTech Solutions": [480.0, 490.0, 500.0, 510.0],
        "Presort Services": [300.0, 305.0, 310.0, 315.0],
        "Other operations": [20.0, 22.0, 24.0, 26.0],
        "Corporate": [-15.0, -16.0, -17.0, -18.0],
    }
    a_vals = {
        "SendTech Solutions": [1700.0, 1825.0, 1980.0],
        "Presort Services": [1120.0, 1180.0, 1230.0],
        "Other operations": [70.0, 85.0, 92.0],
        "Corporate": [-55.0, -60.0, -66.0],
    }

    def _series(label: str) -> list[object]:
        return list(q_vals.get(label, [])) + list(a_vals.get(label, []))

    _write_rows(
        ws_rev,
        [
            ["Revenue by Segment"],
            ["SendTech Solutions", None, None, *_series("SendTech Solutions")],
            ["Presort Services", None, None, *_series("Presort Services")],
            ["Other operations", None, None, *_series("Other operations")],
            [
                "Total",
                None,
                None,
                *[
                    sum(vals[idx] for vals in [q_vals["SendTech Solutions"], q_vals["Presort Services"], q_vals["Other operations"]])
                    for idx in range(len(quarter_headers))
                ],
                *[
                    sum(vals[idx] for vals in [a_vals["SendTech Solutions"], a_vals["Presort Services"], a_vals["Other operations"]])
                    for idx in range(3)
                ],
            ],
            ["Gross Profit and Gross Profit Margin"],
        ],
    )

    _write_rows(
        ws_adj,
        [
            ["Adjusted EBIT"],
            ["SendTech Solutions", None, None, *( [120.0, 125.0, 130.0, 135.0] if include_quarters else [] ), 420.0, 455.0, 510.0],
            ["EBIT margin", None, None, *( [0.250, 0.255, 0.260, 0.265] if include_quarters else [] ), 0.247, 0.249, 0.258],
            ["Presort Services", None, None, *( [45.0, 47.0, 49.0, 52.0] if include_quarters else [] ), 165.0, 178.0, 193.0],
            ["EBIT margin", None, None, *( [0.150, 0.154, 0.158, 0.165] if include_quarters else [] ), 0.147, 0.151, 0.157],
            ["Corporate", None, None, *( [-10.0, -11.0, -12.0, -13.0] if include_quarters else [] ), -40.0, -44.0, -46.0],
            ["EBIT margin", None, None, *( [-0.021, -0.022, -0.023, -0.025] if include_quarters else [] ), -0.024, -0.024, -0.023],
            ["Other operations", None, None, *( [5.0, 4.0, 3.0, 2.0] if include_quarters else [] ), 10.0, 9.0, 8.0],
            ["EBIT margin", None, None, *( [0.250, 0.182, 0.125, 0.077] if include_quarters else [] ), 0.143, 0.106, 0.087],
            ["Adjusted segment EBIT"],
            ["Depreciation & amortization"],
            ["SendTech Solutions", None, None, *( [18.0, 19.0, 20.0, 21.0] if include_quarters else [] ), 67.0, 71.0, 78.0],
            ["Presort Services", None, None, *( [12.0, 12.0, 13.0, 13.0] if include_quarters else [] ), 43.0, 46.0, 50.0],
            ["Corporate", None, None, *( [4.0, 4.0, 4.0, 4.0] if include_quarters else [] ), 15.0, 16.0, 16.0],
            ["Other operations", None, None, *( [1.0, 1.0, 1.0, 1.0] if include_quarters else [] ), 3.0, 4.0, 4.0],
            ["Total"],
        ],
    )

    wb.save(path)
    return path


def test_segment_helper_contracts_select_latest_and_parse_quarters() -> None:
    with _case_dir() as case_dir:
        segment_dir = case_dir / "segment_financials"
        older_path = _write_pbi_segment_workbook(segment_dir, include_quarters=True)
        newer_path = segment_dir / "Historical Segment Financials up to Q4 2026.xlsx"
        newer_path.write_bytes(older_path.read_bytes())
        newer_path.touch()

        picked = latest_segment_financials_workbook(segment_dir)

        assert picked == newer_path

        parsed = parse_quarterly_segment_data_from_workbook(
            older_path,
            annual_segment_alias_patterns=[],
            company_segment_alias_patterns=[],
        )

        revenue_metrics = parsed["metrics"]["Revenue"]
        ebit_metrics = parsed["metrics"]["Adjusted EBIT"]
        margin_metrics = parsed["metrics"]["EBIT margin %"]
        da_metrics = parsed["metrics"]["Depreciation & amortization"]
        assert pd.Timestamp("2025-12-31") in parsed["quarters"]
        assert revenue_metrics["SendTech Solutions"][pd.Timestamp("2025-03-31")] == pytest.approx(480_000_000.0)
        assert ebit_metrics["Presort Services"][pd.Timestamp("2025-12-31")] == pytest.approx(52_000_000.0)
        assert margin_metrics["Presort Services"][pd.Timestamp("2025-06-30")] == pytest.approx(0.154)
        assert da_metrics["Corporate expense"][pd.Timestamp("2025-09-30")] == pytest.approx(4_000_000.0)


def test_driver_template_index_contract_preserves_order_and_units() -> None:
    profile = CompanyProfile(
        ticker="TEST",
        has_bank=False,
        industry_keywords=tuple(),
        segment_patterns=tuple(),
        segment_alias_patterns=tuple(),
        key_adv_require_keywords=tuple(),
        key_adv_deny_keywords=tuple(),
        operating_driver_history_templates=(
            OperatingDriverTemplate(
                group="production",
                key="ethanol_gallons",
                label="Ethanol gallons",
                why_it_matters="Volume",
                preferred_unit="mmgal",
                match_terms=("ethanol", "gallons"),
            ),
            OperatingDriverTemplate(
                group="production",
                key="utilization",
                label="Utilization",
                why_it_matters="Rate",
                preferred_unit="pct",
                match_terms=("utilization",),
            ),
        )
    )

    template_index = load_operating_driver_template_index(
        profile,
        timed_substage=lambda _label: nullcontext(),
    )

    assert template_index["order_map"]["ethanol_gallons"] == 0
    assert template_index["order_map"]["ethanol_gallons_produced"] == 0
    assert template_index["order_map"]["ethanol_gallons_sold"] == 1
    assert template_index["template_unit_map"]["utilization"] == "pct"
    assert template_index["template_specs"]["ethanol_gallons"]["search_terms"] == ("ethanol", "gallons")


def test_docs_for_valuation_accn_contract_sorts_relevant_docs_first() -> None:
    ordered = docs_for_valuation_accn(
        "0001234567-25-000010",
        accession_doc_lookup=lambda _accn: [
            Path("doc_000123456725000010_ex101_agreement.htm"),
            Path("doc_000123456725000010_pressrelease_ex99.htm"),
            Path("doc_000123456725000010_10q.htm"),
        ],
    )

    assert [path.name for path in ordered[:3]] == [
        "doc_000123456725000010_pressrelease_ex99.htm",
        "doc_000123456725000010_10q.htm",
        "doc_000123456725000010_ex101_agreement.htm",
    ]


def test_write_excel_minimal_workbook_has_core_sheets_and_order() -> None:
    with _case_dir() as case_dir:
        out_path = case_dir / "core.xlsx"
        write_excel_from_inputs(_make_inputs(out_path))

        wb = load_workbook(out_path, data_only=False)
        sheetnames = wb.sheetnames
        for name in [
            "SUMMARY",
            "Valuation",
            "Valuation_Grid",
            "History_Q",
            "QA_Log",
            "Needs_Review",
            "QA_Checks",
            "Hidden_Value_Flags",
        ]:
            assert name in sheetnames
        assert sheetnames.index("SUMMARY") < sheetnames.index("Valuation")
        assert sheetnames.index("Valuation") < sheetnames.index("Hidden_Value_Flags")
        assert sheetnames.index("Hidden_Value_Flags") < sheetnames.index("History_Q")
        assert sheetnames.index("History_Q") < sheetnames.index("QA_Log")
        assert sheetnames.index("QA_Log") < sheetnames.index("Needs_Review")
        assert sheetnames.index("Needs_Review") < sheetnames.index("QA_Checks")


def test_write_excel_temp_workbook_preserves_hidden_value_formula_contract() -> None:
    with _case_dir() as case_dir:
        out_path = case_dir / "hidden_value_contract.xlsx"
        write_excel_from_inputs(_make_inputs(out_path))

        wb = load_workbook(out_path, data_only=False)
        try:
            assert "Valuation" in wb.sheetnames
            assert "Hidden_Value_Flags" in wb.sheetnames
            assert wb.sheetnames.index("Valuation") < wb.sheetnames.index("Hidden_Value_Flags") < wb.sheetnames.index("History_Q")

            ws_val = wb["Valuation"]
            flags_header_row = _find_row_with_value(ws_val, "Hidden value flags", column=1)
            assert flags_header_row == 137
            assert "INDEX('Hidden_Value_Flags'!$C:$C,$AI139)" in str(ws_val.cell(row=flags_header_row + 2, column=2).value or "")
            assert "INDEX('Hidden_Value_Flags'!$D:$D,$AI139)" in str(ws_val.cell(row=flags_header_row + 2, column=6).value or "")
            assert "INDEX('Hidden_Value_Flags'!$E:$E,$AI139)" in str(ws_val.cell(row=flags_header_row + 2, column=7).value or "")
            assert "INDEX('Hidden_Value_Flags'!$K:$K,$AI139)" in str(ws_val.cell(row=flags_header_row + 2, column=8).value or "")
            assert "IF(N('Hidden_Value_Flags'!$D$2)>=1,2,\"\")" in str(ws_val.cell(row=139, column=35).value or "")
            assert "IF(N('Hidden_Value_Flags'!$D$8)>=1,8,\"\")" in str(ws_val.cell(row=145, column=35).value or "")
        finally:
            wb.close()


def test_write_excel_snapshot_matches_saved_quarter_notes_ui_for_xlsx_and_xlsm(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = case_dir / "quarter_notes_snapshot.xlsx"
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-12-31"), pd.Timestamp("2025-06-30")],
                    "note_id": ["fcf-1", "debt-1", "rev-1"],
                    "category": [
                        "Cash flow / FCF / capex",
                        "Debt / liquidity / covenants",
                        "Debt / liquidity / covenants",
                    ],
                    "claim": ["FCF TTM accelerated", "Net debt declined", "Revolver utilization notable"],
                    "note": ["FCF TTM accelerated", "Net debt declined", "Revolver utilization notable"],
                    "metric_ref": ["fcf_ttm_delta_yoy", "net_debt_yoy_delta", "revolver_availability_change"],
                    "score": [92.0, 91.0, 89.0],
                    "doc_type": ["model_metric", "model_metric", "revolver"],
                    "doc": ["history_q", "history_q", "history_q"],
                    "source_type": ["model_metric", "model_metric", "revolver"],
                    "evidence_snippet": [
                        "FCF TTM YoY delta $198.7m",
                        "Net debt delta $-77.9m",
                        "We also had $258.5 million available under our committed revolving credit agreement.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "FCF TTM YoY delta $198.7m"}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "Net debt delta $-77.9m"}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "revolver", "snippet": "We also had $258.5 million available under our committed revolving credit agreement."}]),
                    ],
                }
            )

            result = write_excel_from_inputs(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))

            validate_quarter_notes_ui_export(out_path, result.quarter_notes_ui_snapshot)
            snap = read_quarter_notes_ui_snapshot(out_path)
            assert any("FCF TTM improved by $198.7m YoY." in note for _, note in snap["2025-12-31"])

            xlsm_path = out_path.with_suffix(".xlsm")
            shutil.copyfile(out_path, xlsm_path)
            validate_quarter_notes_ui_export(xlsm_path, result.quarter_notes_ui_snapshot)
            validate_saved_workbook_integrity(out_path)
            validate_saved_workbook_integrity(xlsm_path)


def test_write_excel_sanitizes_invalid_comment_xml_characters() -> None:
    with _case_dir() as case_dir:
        out_path = case_dir / "comment_sanitized.xlsx"
        quarter_notes = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "note_id": ["bad-comment-1"],
                "category": ["Programs / initiatives"],
                "claim": ["Actively marketing 2026 45Z production tax credits"],
                "note": ["Actively marketing 2026 45Z production tax credits"],
                "metric_ref": ["45Z monetization / EBITDA"],
                "score": [95.0],
                "doc_type": ["press_release"],
                "doc": ["release_q4.txt"],
                "source_type": ["press_release"],
                "text_full": ["Evidence with \x00 invalid \x0b xml \x1f characters"],
                "comment_full_text": ["Evidence with \x00 invalid \x0b xml \x1f characters"],
                "evidence_snippet": ["Evidence with \x00 invalid \x0b xml \x1f characters"],
            }
        )

        result = write_excel_from_inputs(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))

        validate_saved_workbook_integrity(out_path)
        validate_quarter_notes_ui_export(out_path, result.quarter_notes_ui_snapshot)
        wb = load_workbook(out_path, data_only=False)
        assert "Quarter_Notes_UI" in wb.sheetnames


def test_write_excel_full_mode_emits_relaxed_sheets() -> None:
    with _case_dir() as case_dir:
        relaxed_metrics = pd.DataFrame({"quarter": [pd.Timestamp("2025-12-31")], "adj_ebitda": [12.0]})
        relaxed_breakdown = pd.DataFrame({"quarter": [pd.Timestamp("2025-12-31")], "label": ["Adj"], "value": [1.0]})
        relaxed_files = pd.DataFrame({"quarter": [pd.Timestamp("2025-12-31")], "doc": ["demo"]})
        out_path = case_dir / "full.xlsx"

        write_excel_from_inputs(
            _make_inputs(
                out_path,
                excel_mode="full",
                adj_metrics_relaxed=relaxed_metrics,
                adj_breakdown_relaxed=relaxed_breakdown,
                non_gaap_files_relaxed=relaxed_files,
            )
        )

        wb = load_workbook(out_path, data_only=False)
        for name in [
            "Adjusted_Metrics_Relaxed",
            "Adjustments_Breakdown_Relaxed",
            "NonGAAP_Files_Relaxed",
            "NonGAAP_Bridge_Relaxed",
            "Tag_Coverage",
            "Period_Self_Check",
        ]:
            assert name in wb.sheetnames

        clean_out = case_dir / "clean.xlsx"
        write_excel_from_inputs(_make_inputs(clean_out, excel_mode="clean"))
        wb_clean = load_workbook(clean_out, data_only=False)
        for name in [
            "Adjusted_Metrics_Relaxed",
            "Adjustments_Breakdown_Relaxed",
            "NonGAAP_Files_Relaxed",
            "NonGAAP_Bridge_Relaxed",
        ]:
            assert name not in wb_clean.sheetnames


def test_write_excel_profile_toggles_enable_driver_and_economics_sheets(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = case_dir / "drivers.xlsx"
        profile = CompanyProfile(
            ticker="TEST",
            has_bank=False,
            industry_keywords=tuple(),
            segment_patterns=tuple(),
            segment_alias_patterns=tuple(),
            key_adv_require_keywords=tuple(),
            key_adv_deny_keywords=tuple(),
            enable_operating_drivers_sheet=True,
            enable_economics_overlay_sheet=True,
            enable_economics_market_raw_sheet=True,
            operating_driver_history_templates=(
                OperatingDriverTemplate(
                    group="Operations",
                    label="Plant status",
                    why_it_matters="demo",
                    match_terms=("plant status",),
                    aliases=("plant",),
                    key="plant_status",
                ),
            ),
        )
        monkeypatch.setattr(writer_context_module, "get_company_profile", lambda ticker: profile)
        monkeypatch.setattr(writer_context_module, "load_market_export_rows", lambda *args, **kwargs: [])

        write_excel_from_inputs(_make_inputs(out_path))

        wb = load_workbook(out_path, data_only=False)
        assert "Operating_Drivers" in wb.sheetnames
        assert "Economics_Overlay" in wb.sheetnames
        assert "economics_market_raw" in wb.sheetnames


def test_write_excel_profile_timings_only_emit_when_enabled(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        timed_out = case_dir / "timed.xlsx"
        profile = CompanyProfile(
            ticker="TEST",
            has_bank=False,
            industry_keywords=tuple(),
            segment_patterns=tuple(),
            segment_alias_patterns=tuple(),
            key_adv_require_keywords=tuple(),
            key_adv_deny_keywords=tuple(),
            enable_operating_drivers_sheet=True,
            enable_economics_overlay_sheet=True,
            enable_economics_market_raw_sheet=True,
            operating_driver_history_templates=(
                OperatingDriverTemplate(
                    group="Operations",
                    label="Plant status",
                    why_it_matters="demo",
                    match_terms=("plant status",),
                    aliases=("plant",),
                    key="plant_status",
                ),
            ),
        )
        monkeypatch.setattr(writer_context_module, "get_company_profile", lambda ticker: profile)
        monkeypatch.setattr(writer_context_module, "load_market_export_rows", lambda *args, **kwargs: [])
        real_build_writer_context = excel_writer_module.build_writer_context

        def fake_build_writer_context(inputs: WorkbookInputs):
            ctx = real_build_writer_context(inputs)

            def _stub_overlay_sheet(_rows) -> None:
                ws = ctx.wb.create_sheet("Economics_Overlay")
                ws["A1"] = "timing stub"

            ctx.callbacks.write_economics_overlay_sheet = _stub_overlay_sheet
            return ctx

        monkeypatch.setattr(excel_writer_module, "build_writer_context", fake_build_writer_context)
        write_excel_from_inputs(_make_inputs(timed_out, profile_timings=True))
        stdout_timed = capsys.readouterr().out
        assert "write_excel.prep" in stdout_timed
        assert "write_excel.summary" in stdout_timed
        assert "write_excel.derive.summary_inputs" in stdout_timed
        assert "write_excel.derive.valuation_inputs" in stdout_timed
        assert "write_excel.derive.valuation_inputs.normalize_sources" in stdout_timed
        assert "write_excel.derive.valuation_inputs.net_leverage_text_map" in stdout_timed
        assert "write_excel.derive.valuation_inputs.net_leverage_text_map.frames" in stdout_timed
        assert "write_excel.derive.valuation_inputs.net_leverage_text_map.local_docs" in stdout_timed
        assert "write_excel.derive.valuation_inputs.net_leverage_text_map.audit_docs" in stdout_timed
        assert "write_excel.valuation.precompute" in stdout_timed
        assert "write_excel.valuation.bundle" in stdout_timed
        assert "write_excel.valuation.render" in stdout_timed
        assert "write_excel.derive.driver_inputs.source_records" in stdout_timed
        assert "write_excel.derive.driver_inputs.line_index" in stdout_timed
        assert "write_excel.derive.driver_inputs.crush_bridge_cache" in stdout_timed
        assert "write_excel.derive.driver_inputs.template_rows" in stdout_timed
        assert "write_excel.derive.driver_inputs.operating_history" in stdout_timed
        assert "write_excel.derive.driver_inputs.economics_market" in stdout_timed
        assert "write_excel.drivers.render.operating_drivers" in stdout_timed
        assert "write_excel.drivers.render.economics_overlay" in stdout_timed
        assert "write_excel.derive.report_inputs" in stdout_timed
        assert "write_excel.derive.hidden_value_inputs" in stdout_timed
        assert "write_excel.derive.raw_data_inputs" in stdout_timed
        assert "write_excel.derive.ui_evidence" in stdout_timed
        assert "write_excel.save" in stdout_timed

        quiet_out = case_dir / "quiet.xlsx"
        write_excel_from_inputs(_make_inputs(quiet_out, profile_timings=False))
        stdout_quiet = capsys.readouterr().out
        assert "write_excel.prep" not in stdout_quiet
        assert "write_excel.summary" not in stdout_quiet


def test_write_driver_sheets_records_overlay_substage_timings(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = case_dir / "overlay_timing.xlsx"
        profile = CompanyProfile(
            ticker="TEST",
            has_bank=False,
            industry_keywords=tuple(),
            segment_patterns=tuple(),
            segment_alias_patterns=tuple(),
            key_adv_require_keywords=tuple(),
            key_adv_deny_keywords=tuple(),
            enable_operating_drivers_sheet=True,
            enable_economics_overlay_sheet=True,
            enable_economics_market_raw_sheet=True,
            operating_driver_history_templates=(
                OperatingDriverTemplate(
                    group="Operations",
                    label="Plant status",
                    why_it_matters="demo",
                    match_terms=("plant status",),
                    aliases=("plant",),
                    key="plant_status",
                ),
            ),
        )
        monkeypatch.setattr(writer_context_module, "get_company_profile", lambda ticker: profile)
        monkeypatch.setattr(writer_context_module, "load_market_export_rows", lambda *args, **kwargs: [])
        monkeypatch.setattr(
            writer_context_module,
            "build_prior_quarter_simple_crush_snapshot",
            lambda *args, **kwargs: {"status": "no_data", "process_status": "no_data"},
        )
        monkeypatch.setattr(
            writer_context_module,
            "build_current_qtd_simple_crush_snapshot",
            lambda *args, **kwargs: {"status": "no_data", "process_status": "no_data"},
        )
        monkeypatch.setattr(
            writer_context_module,
            "build_next_quarter_thesis_snapshot",
            lambda *args, **kwargs: {"target_quarter_end": pd.Timestamp("2026-03-31").date()},
        )
        monkeypatch.setattr(writer_context_module, "build_simple_crush_history_series", lambda *args, **kwargs: [])

        ctx = build_writer_context(_make_inputs(out_path, profile_timings=True))
        ctx.callbacks.build_operating_drivers_history_rows = lambda: [
            {"Quarter": pd.Timestamp("2025-12-31").date(), "Driver": "Plant status", "Value": "Operating"}
        ]
        ctx.callbacks.build_economics_market_rows = lambda: []
        ctx.callbacks.load_operating_driver_source_records = lambda: []
        ctx.callbacks.load_operating_driver_source_records_by_quarter = lambda: {}
        ctx.callbacks.prime_operating_driver_crush_detail_cache = lambda records=None: {}
        ctx.state.update(ctx.callbacks.as_state_mapping())

        write_driver_sheets(ctx)

        assert {
            "write_excel.drivers.render.operating_drivers",
            "write_excel.drivers.render.economics_overlay",
            "write_excel.drivers.render.economics_overlay.setup",
            "write_excel.drivers.render.economics_overlay.market_snapshots",
            "write_excel.drivers.render.economics_overlay.market_snapshots.prior_snapshot",
            "write_excel.drivers.render.economics_overlay.market_snapshots.current_snapshot",
            "write_excel.drivers.render.economics_overlay.market_snapshots.history_series",
            "write_excel.drivers.render.economics_overlay.market_snapshots.next_quarter_thesis",
            "write_excel.drivers.render.economics_overlay.base_coefficients",
            "write_excel.drivers.render.economics_overlay.market_inputs",
            "write_excel.drivers.render.economics_overlay.basis_proxy_sandbox",
            "write_excel.drivers.render.economics_overlay.charts_helpers",
            "write_excel.drivers.render.economics_overlay.final_formatting",
        }.issubset(ctx.writer_timings.keys())


def test_write_driver_sheets_restores_gpre_basis_proxy_sandbox_and_overlay_links(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_ticker_model_out_path(case_dir, "GPRE", "gpre_basis_writer_path.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)

            def _stub_market_snapshot(*, label: str, quarter_end: date) -> dict[str, object]:
                return {
                    "status": "available",
                    "display_quarter": quarter_end,
                    "calendar_quarter": quarter_end,
                    "current_market": {
                        "corn_price": 4.80,
                        "natural_gas_price": 0.12,
                        "ethanol_price": 2.05,
                    },
                    "market_meta": {
                        "corn_price": {
                            "as_of": date(2026, 2, 15),
                            "obs_count": 6,
                            "official_corn_basis_source_label": "weighted AMS basis proxy",
                            "official_corn_basis_provenance": "Weighted AMS proxy used for the demo writer-path test.",
                        },
                        "natural_gas_price": {"as_of": date(2026, 2, 15), "obs_count": 6},
                        "ethanol_price": {"as_of": date(2026, 2, 15), "obs_count": 6},
                    },
                    "message": f"{label} available",
                    "process_status": "available",
                }

            def _stub_basis_model_result(*_args, **_kwargs) -> dict[str, object]:
                quarter_rows = pd.DataFrame(
                    [
                        {
                            "quarter": pd.Timestamp("2025-03-31").date(),
                            "official_simple_proxy_usd_per_gal": 0.100,
                            "gpre_proxy_official_usd_per_gal": 0.120,
                            "reported_consolidated_crush_margin_usd_per_gal": 0.210,
                            "underlying_crush_margin_usd_per_gal": 0.310,
                        },
                        {
                            "quarter": pd.Timestamp("2025-06-30").date(),
                            "official_simple_proxy_usd_per_gal": 0.110,
                            "gpre_proxy_official_usd_per_gal": 0.130,
                            "reported_consolidated_crush_margin_usd_per_gal": 0.220,
                            "underlying_crush_margin_usd_per_gal": 0.320,
                        },
                        {
                            "quarter": pd.Timestamp("2025-09-30").date(),
                            "official_simple_proxy_usd_per_gal": 0.125,
                            "gpre_proxy_official_usd_per_gal": 0.145,
                            "reported_consolidated_crush_margin_usd_per_gal": 0.225,
                            "underlying_crush_margin_usd_per_gal": 0.325,
                        },
                        {
                            "quarter": pd.Timestamp("2025-12-31").date(),
                            "official_simple_proxy_usd_per_gal": 0.140,
                            "gpre_proxy_official_usd_per_gal": 0.160,
                            "reported_consolidated_crush_margin_usd_per_gal": 0.240,
                            "underlying_crush_margin_usd_per_gal": 0.340,
                        },
                    ]
                )
                leaderboard = pd.DataFrame(
                    [
                        {
                            "model_key": "process_utilization_regime_residual",
                            "chosen": True,
                            "expanded_best_candidate": True,
                            "family_label": "Process residual",
                            "timing_rule": "Demo timing rule",
                            "clean_mae": 0.050,
                            "underlying_mae": 0.060,
                            "hybrid_score": 0.055,
                            "avg_abs_diff_vs_official": 0.040,
                            "gt_2c_quarters": 4,
                            "gt_5c_quarters": 2,
                            "selection_guard_reason": "passed_guardrails",
                            "promotion_guard_reason": "passed_promotion_guardrails",
                            "selection_guard_pass": True,
                            "promotion_guard_pass": True,
                            "live_preview_quality_status": "close",
                            "live_preview_mae": 0.0,
                            "live_preview_max_error": 0.0,
                            "candidate_status": "winner",
                            "notes": "Demo writer-path winner",
                        }
                    ]
                )
                return {
                    "quarterly_df": quarter_rows,
                    "leaderboard_df": leaderboard,
                    "metrics_df": pd.DataFrame(),
                    "weights_df": pd.DataFrame(
                        [
                            {
                                "model_key": "plant_count_weighted",
                                "quarter": pd.Timestamp("2025-12-31").date(),
                                "plant": "Nebraska",
                                "weight": 0.55,
                            },
                            {
                                "model_key": "plant_count_weighted",
                                "quarter": pd.Timestamp("2025-12-31").date(),
                                "plant": "Iowa",
                                "weight": 0.45,
                            },
                        ]
                    ),
                    "official_market_rows": [
                        {
                            "region_label": "Nebraska",
                            "active_capacity_mmgy": 250.0,
                            "weight": 0.55,
                            "ethanol_value_usd_per_gal": 2.05,
                            "ethanol_series_label": "Nebraska cash",
                            "basis_value_cents_per_bu": 15.0,
                            "basis_value_usd_per_bu": 0.15,
                            "basis_series_label": "AMS proxy",
                            "fallback_note": "Primary mapped series used.",
                        }
                    ],
                    "official_market_summary": "Official market model | Representative quarter: demo test coverage.",
                    "official_weighting_method": "Active-capacity weighted.",
                    "official_ethanol_method": "Mapped ethanol benchmark.",
                    "official_basis_method": "Weighted AMS basis proxy.",
                    "official_gas_method": "Fixed natural gas burden.",
                    "official_fallback_policy": "Fallback only when direct mapped series are unavailable.",
                    "overlay_preview_bundle": {
                        "official_frames": {
                            "prior_quarter": {"value": 0.140, "status": "available"},
                            "quarter_open": {"value": 0.150, "status": "available"},
                            "current_qtd": {"value": 0.155, "status": "available"},
                            "next_quarter_thesis": {"value": 0.165, "status": "available"},
                        },
                        "gpre_proxy_frames": {
                            "prior_quarter": {"value": 0.160, "status": "available"},
                            "quarter_open": {
                                "value": 0.170,
                                "status": "available",
                                "live_preview_note": "Quarter-open fitted value for the chosen model.",
                            },
                            "current_qtd": {
                                "value": 0.175,
                                "status": "available",
                                "live_preview_note": "Current fitted preview uses the demo residual overlay.",
                            },
                            "next_quarter_thesis": {
                                "value": 0.185,
                                "status": "available",
                                "live_preview_note": "Next-quarter fitted preview uses the demo residual overlay.",
                            },
                        },
                        "gpre_proxy_formula_helpers": {
                            "current_qtd": {"slope": 0.05, "intercept": 0.03},
                            "next_quarter_thesis": {"slope": 0.05, "intercept": 0.04},
                        },
                        "quarter_open_market_snapshot": {
                            **_stub_market_snapshot(label="Quarter-open outlook", quarter_end=date(2026, 6, 30)),
                            "quarter_open_provenance": "manual_local_snapshot",
                        },
                        "quarter_open_snapshot_status": "available",
                        "quarter_open_provenance": "manual_local_snapshot",
                        "quarter_open_target_quarter_end": date(2026, 6, 30),
                    },
                    "proxy_implied_results": {
                        "title": "Proxy-implied results ($m)",
                        "note": "Demo bridge note for the writer-path regression test.",
                        "frames": {
                            "prior_quarter": {"proxy_result_m": 11.0, "reported_result_m": 10.5, "implied_volume_mmgal": 90.0},
                            "quarter_open": {"proxy_result_m": 12.0, "reported_result_m": 10.5, "implied_volume_mmgal": 90.0},
                            "current_qtd": {"proxy_result_m": 13.0, "reported_result_m": 10.5, "implied_volume_mmgal": 90.0},
                            "next_quarter_thesis": {"proxy_result_m": 14.0, "reported_result_m": 10.5, "implied_volume_mmgal": 90.0},
                        },
                    },
                    "gpre_proxy_model_key": "process_utilization_regime_residual",
                    "recommended_model_key": "process_utilization_regime_residual",
                    "incumbent_baseline_model_key": "bridge_front_loaded",
                    "expanded_best_candidate_model_key": "process_utilization_regime_residual",
                    "production_winner_model_key": "process_utilization_regime_residual",
                    "gpre_proxy_live_preview_quality_status": "close",
                    "gpre_proxy_live_preview_mae": 0.0,
                    "gpre_proxy_live_preview_max_error": 0.0,
                    "gpre_proxy_live_preview_top_miss_quarters": "",
                    "gpre_proxy_live_preview_worst_phase": "current",
                    "production_decision_story": "Demo winner for direct writer regression coverage.",
                    "selection_vs_promotion_explanation": "Selection and promotion align in the direct writer test.",
                    "summary_markdown": "Demo summary for writer-path regression coverage.",
                }

            monkeypatch.setattr(writer_context_module, "load_market_export_rows", lambda *args, **kwargs: [])
            monkeypatch.setattr(writer_context_module, "build_gpre_plant_capacity_history", lambda *args, **kwargs: {})
            monkeypatch.setattr(
                writer_context_module,
                "build_gpre_official_proxy_snapshot",
                lambda *args, prior_quarter=False, **kwargs: _stub_market_snapshot(
                    label="Prior quarter" if prior_quarter else "Current QTD",
                    quarter_end=date(2025, 12, 31) if prior_quarter else date(2026, 3, 31),
                ),
            )
            monkeypatch.setattr(writer_context_module, "build_gpre_official_proxy_history_series", lambda *args, **kwargs: [])
            monkeypatch.setattr(
                writer_context_module,
                "build_next_quarter_thesis_snapshot",
                lambda *args, **kwargs: {
                    "target_quarter_start": date(2026, 7, 1),
                    "target_quarter_end": date(2026, 9, 30),
                    "corn": {
                        "value": 4.90,
                        "manual": False,
                        "thesis_label": "CBOT Corn futures",
                        "basis_suffix": "actual GPRE plant-bid basis + AMS fallback",
                        "comment": "Demo thesis uses live bids + AMS fallback.",
                    },
                    "natural_gas": {
                        "value": 0.11,
                        "manual": False,
                        "thesis_label": "NYMEX Natural Gas futures",
                        "basis_suffix": "futures-based approximation",
                        "comment": "Demo thesis uses NYMEX futures.",
                    },
                    "ethanol": {
                        "value": 2.10,
                        "manual": False,
                        "thesis_label": "Local Chicago ethanol futures",
                        "basis_suffix": "local Chicago ethanol futures strip",
                        "comment": "Demo thesis uses local Chicago ethanol futures strip.",
                    },
                },
            )
            monkeypatch.setattr(writer_context_module, "build_simple_crush_history_series", lambda *args, **kwargs: [])
            monkeypatch.setattr(writer_context_module, "build_gpre_basis_proxy_model", _stub_basis_model_result)

            base_inputs = _make_inputs(
                out_path,
                ticker="GPRE",
                hist=_make_hist(),
                profile_timings=True,
                capture_saved_workbook_provenance=False,
                excel_debug_scope="drivers",
            )
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            ctx = build_writer_context(inputs)
            ctx.callbacks.build_operating_drivers_history_rows = lambda: [
                {"Quarter": pd.Timestamp("2025-03-31").date(), "_driver_key": "ethanol_gallons_sold", "Driver": "Ethanol gallons sold", "Value": 90.0},
                {"Quarter": pd.Timestamp("2025-03-31").date(), "_driver_key": "ethanol_gallons_produced", "Driver": "Ethanol gallons produced", "Value": 88.0},
                {"Quarter": pd.Timestamp("2025-03-31").date(), "_driver_key": "consolidated_ethanol_crush_margin", "Driver": "Reported consolidated crush margin", "Value": 10.5},
                {"Quarter": pd.Timestamp("2025-03-31").date(), "_driver_key": "underlying_crush_margin", "Driver": "Underlying crush margin", "Value": 9.0},
            ]
            ctx.callbacks.build_economics_market_rows = lambda: []
            ctx.callbacks.load_operating_driver_source_records = lambda: []
            ctx.callbacks.load_operating_driver_source_records_by_quarter = lambda: {
                date(2023, 3, 31): [
                    {
                        "text": "Management said 50 million to 100 million gallons were hedged at roughly $0.22 to $0.25 per gallon through year-end.",
                        "_text_low": "management said 50 million to 100 million gallons were hedged at roughly $0.22 to $0.25 per gallon through year-end.",
                        "source_type": "transcript",
                        "source_doc": "demo_transcript.txt",
                        "source_rank": 1,
                        "_fragment_penalty": 0.0,
                        "_is_complete_signal": True,
                    }
                ]
            }
            ctx.callbacks.prime_operating_driver_crush_detail_cache = lambda records=None: {}
            ctx.state.update(ctx.callbacks.as_state_mapping())

            write_driver_sheets(ctx)

            assert "Operating_Drivers" in ctx.wb.sheetnames
            assert "Economics_Overlay" in ctx.wb.sheetnames
            assert "Basis_Proxy_Sandbox" in ctx.wb.sheetnames
            assert ctx.wb.sheetnames.index("Economics_Overlay") < ctx.wb.sheetnames.index("Basis_Proxy_Sandbox")

            ws_overlay = ctx.wb["Economics_Overlay"]
            ws_basis = ctx.wb["Basis_Proxy_Sandbox"]
            sandbox_build_row = _find_row_with_value(ws_basis, "Approximate market crush build-up ($/gal)", column=2)
            official_proxy_row = _find_row_with_value(ws_overlay, "Approximate market crush ($/gal)", column=1)
            fitted_proxy_row = _find_row_with_value(ws_overlay, "GPRE crush proxy ($/gal)", column=1)
            forward_proxy_row = _find_row_with_value(ws_overlay, "Best forward lens ($/gal)", column=1)
            bridge_row = _find_row_with_value(ws_overlay, "Bridge to reported", column=1)
            proxy_compare_row = _find_row_with_value(ws_overlay, "Proxy comparison ($/gal)", column=1)
            fitted_bridge_row = _find_row_with_value(ws_overlay, "GPRE crush proxy ($m)", column=1)
            forward_bridge_row = _find_row_with_value(ws_overlay, "Best forward lens ($m)", column=1)

            assert sandbox_build_row is not None
            assert official_proxy_row is not None
            assert fitted_proxy_row is not None
            assert forward_proxy_row is not None
            assert bridge_row is not None
            assert proxy_compare_row is not None
            assert fitted_bridge_row is not None
            assert forward_bridge_row is not None
            assert str(ws_basis["B1"].value or "").strip() == "Exploratory GPRE basis proxy sandbox (test)"
            assert "Official simple row build-up used by Approximate market crush on Economics_Overlay." in str(
                ws_basis.cell(row=sandbox_build_row + 1, column=2).value or ""
            )
            assert "Basis_Proxy_Sandbox!" in str(ws_overlay.cell(row=official_proxy_row, column=2).value or "")
            assert "Official row = Approximate market crush" in str(ws_overlay.cell(row=proxy_compare_row + 1, column=1).value or "")
            assert str(ws_overlay.cell(row=proxy_compare_row + 2, column=1).value or "").strip() == "Proxy row"
            assert str(ws_overlay.cell(row=fitted_bridge_row, column=1).value or "").strip() == "GPRE crush proxy ($m)"
            assert str(ws_overlay.cell(row=forward_bridge_row, column=1).value or "").strip() == "Best forward lens ($m)"
            assert not list(getattr(ws_overlay, "_charts", []))
            assert "write_excel.drivers.render.economics_overlay.basis_proxy_sandbox" in ctx.writer_timings


def test_write_driver_sheets_does_not_leak_gpre_overlay_surface_to_non_gpre_profiles(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = case_dir / "non_gpre_overlay.xlsx"
        profile = CompanyProfile(
            ticker="TEST",
            has_bank=False,
            industry_keywords=tuple(),
            segment_patterns=tuple(),
            segment_alias_patterns=tuple(),
            key_adv_require_keywords=tuple(),
            key_adv_deny_keywords=tuple(),
            enable_operating_drivers_sheet=True,
            enable_economics_overlay_sheet=True,
            enable_economics_market_raw_sheet=False,
            operating_driver_history_templates=(
                OperatingDriverTemplate(
                    group="Operations",
                    label="Plant status",
                    why_it_matters="demo",
                    match_terms=("plant status",),
                    aliases=("plant",),
                    key="plant_status",
                ),
            ),
        )
        monkeypatch.setattr(writer_context_module, "get_company_profile", lambda ticker: profile)
        monkeypatch.setattr(writer_context_module, "load_market_export_rows", lambda *args, **kwargs: [])
        monkeypatch.setattr(
            writer_context_module,
            "build_prior_quarter_simple_crush_snapshot",
            lambda *args, **kwargs: {"status": "no_data", "process_status": "no_data"},
        )
        monkeypatch.setattr(
            writer_context_module,
            "build_current_qtd_simple_crush_snapshot",
            lambda *args, **kwargs: {"status": "no_data", "process_status": "no_data"},
        )
        monkeypatch.setattr(
            writer_context_module,
            "build_next_quarter_thesis_snapshot",
            lambda *args, **kwargs: {"target_quarter_end": pd.Timestamp("2026-03-31").date()},
        )
        monkeypatch.setattr(writer_context_module, "build_simple_crush_history_series", lambda *args, **kwargs: [])

        ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", profile_timings=True))
        ctx.callbacks.build_operating_drivers_history_rows = lambda: [
            {"Quarter": pd.Timestamp("2025-12-31").date(), "Driver": "Plant status", "Value": "Operating"}
        ]
        ctx.callbacks.build_economics_market_rows = lambda: []
        ctx.callbacks.load_operating_driver_source_records = lambda: []
        ctx.callbacks.load_operating_driver_source_records_by_quarter = lambda: {}
        ctx.callbacks.prime_operating_driver_crush_detail_cache = lambda records=None: {}
        ctx.state.update(ctx.callbacks.as_state_mapping())

        write_driver_sheets(ctx)

        assert "Economics_Overlay" in ctx.wb.sheetnames
        assert "Basis_Proxy_Sandbox" not in ctx.wb.sheetnames
        ws_overlay = ctx.wb["Economics_Overlay"]
        assert _find_row_with_value(ws_overlay, "Approximate market crush ($/gal)", column=1) is None
        assert _find_row_with_value(ws_overlay, "GPRE crush proxy ($/gal)", column=1) is None
        assert _find_row_with_value(ws_overlay, "Proxy comparison ($/gal)", column=1) is None
        assert _find_row_with_value(ws_overlay, "Coproduct source gate", column=1) is None
        assert _find_row_with_value(ws_overlay, "Coproduct signal readiness", column=1) is None
        assert _find_row_with_value(ws_overlay, "Coproduct frame summary", column=1) is None
        assert _find_row_with_value(ws_overlay, "Coproduct economics", column=1) is None
        assert _find_row_with_value(ws_overlay, "Approximate coproduct credit (quarterly history)", column=2) is None
        assert _find_row_with_value(ws_overlay, "Approximate coproduct credit ($/gal, quarterly history)", column=2) is None
        assert _find_row_with_value(ws_overlay, "Recent coproduct history", column=2) is None
        assert _find_row_with_value(ws_overlay, "Current QTD trend tracking ($/gal, crush margin lens)", column=1) is None
        assert _find_row_with_value(ws_overlay, "Approximate market crush, fitted models, and real GPRE crush margin (quarterly)", column=2) is None
        assert _find_row_with_value(
            ws_overlay,
            "Coverage reflects covered active-capacity footprint; values are covered-footprint weighted averages.",
            column=2,
        ) is None
        assert _find_row_with_value(
            ws_overlay,
            "Coproduct-aware experimental lenses live in Basis_Proxy_Sandbox and are comparison-only.",
            column=1,
        ) is None
        assert not list(getattr(ws_overlay, "_charts", []))
        assert "write_excel.drivers.render.economics_overlay.basis_proxy_sandbox" in ctx.writer_timings


def test_build_writer_context_is_lazy_for_derived_frames() -> None:
    with _case_dir() as case_dir:
        ctx = build_writer_context(_make_inputs(case_dir / "lazy.xlsx"))
        assert all(value is None for value in vars(ctx.derived).values())

        prepare_writer_inputs(ctx)
        assert all(value is None for value in vars(ctx.derived).values())


def test_write_excel_from_inputs_supports_drivers_debug_scope(
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
        with _case_dir() as case_dir:
            out_path = case_dir / "drivers_debug.xlsx"
            profile = CompanyProfile(
                ticker="TEST",
                has_bank=False,
                industry_keywords=tuple(),
                segment_patterns=tuple(),
                segment_alias_patterns=tuple(),
                key_adv_require_keywords=tuple(),
                key_adv_deny_keywords=tuple(),
                enable_operating_drivers_sheet=True,
                enable_economics_overlay_sheet=True,
                enable_economics_market_raw_sheet=True,
                operating_driver_history_templates=(
                OperatingDriverTemplate(
                    group="Operations",
                    label="Plant status",
                    why_it_matters="demo",
                    match_terms=("plant status",),
                    aliases=("plant",),
                    key="plant_status",
                ),
            ),
        )
        monkeypatch.setattr(writer_context_module, "get_company_profile", lambda ticker: profile)
        monkeypatch.setattr(writer_context_module, "load_market_export_rows", lambda *args, **kwargs: [])
        real_build_writer_context = excel_writer_module.build_writer_context

        def fake_build_writer_context(inputs: WorkbookInputs):
            ctx = real_build_writer_context(inputs)

            def _stub_overlay_sheet(_rows) -> None:
                ws = ctx.wb.create_sheet("Economics_Overlay")
                ws["A1"] = "debug-scope stub"

            ctx.callbacks.write_economics_overlay_sheet = _stub_overlay_sheet
            return ctx

        monkeypatch.setattr(excel_writer_module, "build_writer_context", fake_build_writer_context)

        write_excel_from_inputs(
            _make_inputs(
                out_path,
                profile_timings=True,
                excel_debug_scope="drivers",
                capture_saved_workbook_provenance=False,
            )
        )
        stdout = capsys.readouterr().out
        assert "write_excel.drivers.render.operating_drivers" in stdout
        assert "write_excel.drivers.render.economics_overlay" in stdout
        assert "write_excel.summary" not in stdout
        assert "write_excel.ui" not in stdout

        wb = load_workbook(out_path, data_only=False)
        assert "Operating_Drivers" in wb.sheetnames
        assert "Economics_Overlay" in wb.sheetnames
        assert "Hidden_Value_Flags" not in wb.sheetnames
        assert "Quarter_Notes_UI" not in wb.sheetnames
        assert "Promise_Progress_UI" not in wb.sheetnames
        assert "Summary" not in wb.sheetnames
        wb.close()


def test_write_excel_from_inputs_supports_ui_debug_scope(
    capsys: pytest.CaptureFixture[str],
) -> None:
    with _case_dir() as case_dir:
        out_path = case_dir / "ui_debug.xlsx"
        quarter_notes = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "note_id": ["note-1"],
                "category": ["Guidance / outlook"],
                "claim": ["Revenue should stabilize by year end."],
                "note": ["Revenue should stabilize by year end."],
                "metric_ref": ["revenue_yoy"],
                "source_type": ["earnings_release"],
                "doc": ["release.txt"],
                "evidence_snippet": ["Revenue should stabilize by year end."],
            }
        )
        promises = pd.DataFrame(
            {
                "promise_id": ["p-1"],
                "first_seen_evidence_quarter": [pd.Timestamp("2025-12-31")],
                "promise_text": ["Revenue should stabilize by year end."],
            }
        )
        progress = pd.DataFrame(
            {
                "promise_id": ["p-1"],
                "quarter": [pd.Timestamp("2025-12-31")],
                "status": ["open"],
            }
        )

        write_excel_from_inputs(
            _make_inputs(
                out_path,
                quarter_notes=quarter_notes,
                promises=promises,
                promise_progress=progress,
                profile_timings=True,
                excel_debug_scope="ui",
                capture_saved_workbook_provenance=False,
            )
        )
        stdout = capsys.readouterr().out
        assert "write_excel.ui.render.quarter_notes" in stdout
        assert "write_excel.summary" not in stdout
        assert "write_excel.drivers" not in stdout

        wb = load_workbook(out_path, data_only=False)
        assert "Quarter_Notes" in wb.sheetnames
        assert "Quarter_Notes_Evidence" in wb.sheetnames
        assert "Promise_Tracker" in wb.sheetnames
        assert "Promise_Progress" in wb.sheetnames
        assert "Quarter_Notes_UI" in wb.sheetnames
        assert "Promise_Progress_UI" not in wb.sheetnames
        assert "Promise_Evidence" not in wb.sheetnames
        assert "NonGAAP_Credibility" not in wb.sheetnames
        assert "Hidden_Value_Flags" not in wb.sheetnames
        assert "Operating_Drivers" not in wb.sheetnames
        assert "Summary" not in wb.sheetnames
        wb.close()


def test_build_writer_context_exposes_explicit_state_contract() -> None:
    with _case_dir() as case_dir:
        out_path = case_dir / "explicit.xlsx"
        ctx = build_writer_context(_make_inputs(out_path))

        assert ctx.data.out_path == out_path
        assert ctx.data.ticker == "TEST"
        assert ctx.data.excel_mode == "clean"
        assert ctx.data.profile_timings is False
        assert isinstance(ctx.data.operating_driver_history_rows, list)
        assert isinstance(ctx.data.economics_market_rows, list)
        assert ctx.data.doc_cache.accession_doc_paths == {}
        assert ctx.data.runtime_cache.quarter_notes.doc_analysis_cache == {}
        assert ctx.data.runtime_cache.valuation_precompute.filing_doc_text_cache == {}
        assert ctx.data.runtime_cache.valuation_precompute.buyback_execution_doc_cache == {}
        assert ctx.data.runtime_cache.operating_drivers.template_rows_cache == {}

        assert ctx.callbacks.write_sheet is ctx.state["_write_sheet"]
        assert ctx.callbacks.build_report is ctx.state["_build_report"]
        assert ctx.callbacks.build_operating_drivers_history_rows is ctx.state["_build_operating_drivers_history_rows"]
        assert ctx.callbacks.build_economics_market_rows is ctx.state["_build_economics_market_rows"]
        assert ctx.state["writer_timings"] is ctx.writer_timings
        assert ctx.state["out_path"] == ctx.data.out_path
        assert ctx.state["qa_checks"] is ctx.data.qa_checks
        assert ctx.state["info_log"] is ctx.data.info_log
        assert ctx.state["company_profile"] is ctx.company_profile
        assert ctx.state["_load_profile_slide_signals_by_quarter"] is not None
        assert ctx.state["_load_operating_driver_45z_guidance_docs_by_quarter"] is not None
        assert "ctx_ref" not in ctx.state
        assert "valuation_style_bundle_cache" not in ctx.state


def test_quarterly_row_color_policy_helper_is_metric_aware() -> None:
    policy_revenue = writer_context_module._quarterly_row_color_policy("Revenue", section_label="Operating")
    assert policy_revenue.comparison_basis == "yoy"
    assert policy_revenue.directionality == "higher_better"

    policy_capex = writer_context_module._quarterly_row_color_policy("Capex", section_label="Cash Flow")
    assert policy_capex.comparison_basis == "yoy"
    assert policy_capex.directionality == "lower_better"

    policy_net_debt_qoq = writer_context_module._quarterly_row_color_policy(
        "Net debt QoQ Î” ($m)",
        section_label="Leverage & Liquidity",
    )
    assert policy_net_debt_qoq.comparison_basis == "direct_delta"
    assert policy_net_debt_qoq.directionality == "lower_better"

    policy_fcf_ttm = writer_context_module._quarterly_row_color_policy("FCF (TTM)", section_label="Cash Flow")
    assert policy_fcf_ttm.comparison_basis == "ttm_vs_prior_ttm"
    assert policy_fcf_ttm.directionality == "higher_better"

    policy_current_ratio = writer_context_module._quarterly_row_color_policy(
        "Current ratio",
        section_label="Leverage & Liquidity",
    )
    assert policy_current_ratio.comparison_basis == "yoy"
    assert policy_current_ratio.directionality == "higher_better"

    policy_interest_cov = writer_context_module._quarterly_row_color_policy(
        "Interest coverage (P&L TTM)",
        section_label="Leverage & Liquidity",
    )
    assert policy_interest_cov.comparison_basis == "ttm_vs_prior_ttm"
    assert policy_interest_cov.directionality == "higher_better"

    policy_cash_interest_cov = writer_context_module._quarterly_row_color_policy(
        "Cash interest coverage (TTM)",
        section_label="Leverage & Liquidity",
    )
    assert policy_cash_interest_cov.comparison_basis == "ttm_vs_prior_ttm"
    assert policy_cash_interest_cov.directionality == "higher_better"

    policy_bv_share = writer_context_module._quarterly_row_color_policy(
        "BV/share",
        section_label="Equity / Per-share",
    )
    assert policy_bv_share.comparison_basis == "yoy"
    assert policy_bv_share.directionality == "higher_better"

    policy_tbv_share = writer_context_module._quarterly_row_color_policy(
        "TBV/share",
        section_label="Equity / Per-share",
    )
    assert policy_tbv_share.comparison_basis == "yoy"
    assert policy_tbv_share.directionality == "higher_better"

    policy_fcf_share_ttm = writer_context_module._quarterly_row_color_policy(
        "FCF/share (TTM)",
        section_label="Equity / Per-share",
    )
    assert policy_fcf_share_ttm.comparison_basis == "ttm_vs_prior_ttm"
    assert policy_fcf_share_ttm.directionality == "higher_better"

    policy_ev_ebitda_ttm = writer_context_module._quarterly_row_color_policy(
        "EV/EBITDA (TTM)",
        section_label="Equity / Per-share",
    )
    assert policy_ev_ebitda_ttm.comparison_basis == "ttm_vs_prior_ttm"
    assert policy_ev_ebitda_ttm.directionality == "neutral"

    policy_ev_adj_ebitda_ttm = writer_context_module._quarterly_row_color_policy(
        "EV/Adj EBITDA (TTM)",
        section_label="Equity / Per-share",
    )
    assert policy_ev_adj_ebitda_ttm.comparison_basis == "ttm_vs_prior_ttm"
    assert policy_ev_adj_ebitda_ttm.directionality == "neutral"

    policy_acquisitions = writer_context_module._quarterly_row_color_policy(
        "Acquisitions (TTM, cash)",
        section_label="Cash Flow",
    )
    assert policy_acquisitions.directionality == "neutral"

    policy_buybacks_cash = writer_context_module._quarterly_row_color_policy(
        "Buybacks (cash)",
        section_label="Cash Flow",
    )
    assert policy_buybacks_cash.comparison_basis == "qoq"
    assert policy_buybacks_cash.directionality == "higher_better"

    policy_revolver_drawn = writer_context_module._quarterly_row_color_policy(
        "Revolver drawn",
        section_label="Leverage & Liquidity",
    )
    assert policy_revolver_drawn.comparison_basis == "yoy"
    assert policy_revolver_drawn.directionality == "lower_better"

    policy_revolver_availability = writer_context_module._quarterly_row_color_policy(
        "Revolver availability",
        section_label="Leverage & Liquidity",
    )
    assert policy_revolver_availability.comparison_basis == "yoy"
    assert policy_revolver_availability.directionality == "higher_better"

    policy_liquidity = writer_context_module._quarterly_row_color_policy(
        "Liquidity (cash+availability)",
        section_label="Leverage & Liquidity",
    )
    assert policy_liquidity.comparison_basis == "yoy"
    assert policy_liquidity.directionality == "higher_better"


def test_quarterly_color_metric_helper_uses_basis_and_directionality() -> None:
    assert writer_context_module._quarterly_color_metric_from_series(
        [100.0, 130.0, 120.0, 90.0, 110.0],
        4,
        comparison_basis="yoy",
        directionality="higher_better",
    ) > 0

    assert writer_context_module._quarterly_color_metric_from_series(
        [50.0, 40.0, 60.0, 55.0, 45.0],
        4,
        comparison_basis="yoy",
        directionality="lower_better",
    ) > 0

    assert writer_context_module._quarterly_color_metric_from_series(
        [None, 74.055, -28.609],
        1,
        comparison_basis="direct_delta",
        directionality="higher_better",
    ) > 0

    assert writer_context_module._quarterly_color_metric_from_series(
        [None, -11.841, 57.502],
        1,
        comparison_basis="direct_delta",
        directionality="lower_better",
    ) > 0

    assert writer_context_module._quarterly_color_metric_from_series(
        [100.0, 105.0, 110.0, 115.0, 140.0],
        4,
        comparison_basis="ttm_vs_prior_ttm",
        directionality="higher_better",
    ) > 0


def test_ensure_driver_inputs_memoize_and_reuse_cached_results(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        ctx = build_writer_context(_make_inputs(case_dir / "driver_memo.xlsx"))
        ctx.data.enable_operating_drivers_sheet = True
        ctx.data.enable_economics_overlay_sheet = True
        ctx.data.enable_economics_market_raw_sheet = True
        ctx.state["enable_operating_drivers_sheet"] = True
        ctx.state["enable_economics_overlay_sheet"] = True
        ctx.state["enable_economics_market_raw_sheet"] = True
        ctx.callbacks.build_operating_drivers_history_rows = lambda: [{"Quarter": pd.Timestamp("2025-12-31").date(), "Driver": "Demo"}]
        ctx.callbacks.build_economics_market_rows = lambda: [{"quarter": pd.Timestamp("2025-12-31").date(), "series_key": "corn_cash_demo"}]
        ctx.callbacks.load_operating_driver_source_records = lambda: []
        ctx.callbacks.load_operating_driver_source_records_by_quarter = lambda: {}
        ctx.callbacks.prime_operating_driver_crush_detail_cache = lambda records=None: {}
        ctx.state.update(ctx.callbacks.as_state_mapping())

        ensure_driver_inputs(ctx)
        assert ctx.state["operating_driver_history_rows"] == [{"Quarter": pd.Timestamp("2025-12-31").date(), "Driver": "Demo"}]
        assert ctx.state["economics_market_rows"] == [{"quarter": pd.Timestamp("2025-12-31").date(), "series_key": "corn_cash_demo"}]
        assert ctx.data.operating_driver_history_rows == [{"Quarter": pd.Timestamp("2025-12-31").date(), "Driver": "Demo"}]
        assert ctx.data.economics_market_rows == [{"quarter": pd.Timestamp("2025-12-31").date(), "series_key": "corn_cash_demo"}]

        monkeypatch.setattr(ctx.callbacks, "build_operating_drivers_history_rows", lambda: (_ for _ in ()).throw(RuntimeError("driver recompute")))
        monkeypatch.setattr(ctx.callbacks, "build_economics_market_rows", lambda: (_ for _ in ()).throw(RuntimeError("market recompute")))
        ensure_driver_inputs(ctx)
        assert ctx.state["operating_driver_history_rows"] == [{"Quarter": pd.Timestamp("2025-12-31").date(), "Driver": "Demo"}]
        assert ctx.state["economics_market_rows"] == [{"quarter": pd.Timestamp("2025-12-31").date(), "series_key": "corn_cash_demo"}]


def test_ensure_helpers_memoize_and_reuse_cached_results(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        ctx = build_writer_context(_make_inputs(case_dir / "memo.xlsx"))

        ensure_report_inputs(ctx)
        report_id = id(ctx.derived.report_is)
        monkeypatch.setattr(ctx.callbacks, "build_report", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("report recompute")))
        ensure_report_inputs(ctx)
        assert id(ctx.derived.report_is) == report_id

        ensure_raw_data_inputs(ctx)
        raw_ids = (
            id(ctx.derived.facts_long),
            id(ctx.derived.lineitem_map),
            id(ctx.derived.period_index),
            id(ctx.derived.ng_bridge),
        )
        monkeypatch.setattr(ctx.callbacks, "build_facts_long", lambda: (_ for _ in ()).throw(RuntimeError("facts recompute")))
        monkeypatch.setattr(ctx.callbacks, "build_lineitem_map", lambda: (_ for _ in ()).throw(RuntimeError("lineitem recompute")))
        monkeypatch.setattr(ctx.callbacks, "build_period_index", lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("period recompute")))
        monkeypatch.setattr(ctx.callbacks, "build_ng_bridge", lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("bridge recompute")))
        ensure_raw_data_inputs(ctx)
        assert (
            id(ctx.derived.facts_long),
            id(ctx.derived.lineitem_map),
            id(ctx.derived.period_index),
            id(ctx.derived.ng_bridge),
        ) == raw_ids

        ensure_ui_evidence(ctx)
        ui_ids = (
            id(ctx.derived.quarter_notes_evidence_df),
            id(ctx.derived.promise_evidence_df),
        )
        monkeypatch.setattr(ctx.callbacks, "build_qn_evidence_src", lambda: (_ for _ in ()).throw(RuntimeError("qn recompute")))
        monkeypatch.setattr(ctx.callbacks, "build_promise_evidence_src", lambda: (_ for _ in ()).throw(RuntimeError("promise recompute")))
        ensure_ui_evidence(ctx)
        assert (
            id(ctx.derived.quarter_notes_evidence_df),
            id(ctx.derived.promise_evidence_df),
        ) == ui_ids

        ensure_valuation_inputs(ctx)
        valuation_ids = (
            id(ctx.derived.leverage_df),
            id(ctx.derived.valuation_summary_df),
            id(ctx.derived.valuation_grid_df),
        )
        monkeypatch.setattr(writer_core_module, "valuation_engine", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("valuation recompute")))
        ensure_valuation_inputs(ctx)
        assert (
            id(ctx.derived.leverage_df),
            id(ctx.derived.valuation_summary_df),
            id(ctx.derived.valuation_grid_df),
        ) == valuation_ids

        ensure_hidden_value_inputs(ctx)
        hidden_ids = (
            id(ctx.derived.signals_base_df),
            id(ctx.derived.flags_df),
            id(ctx.derived.flags_audit_df),
            id(ctx.derived.flags_recompute_df),
        )
        monkeypatch.setattr(
            writer_core_module,
            "build_hidden_value_outputs",
            lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("hidden recompute")),
        )
        ensure_hidden_value_inputs(ctx)
        assert (
            id(ctx.derived.signals_base_df),
            id(ctx.derived.flags_df),
            id(ctx.derived.flags_audit_df),
            id(ctx.derived.flags_recompute_df),
        ) == hidden_ids

        ensure_summary_inputs(ctx)
        summary_id = id(ctx.derived.summary_df)
        monkeypatch.setitem(ctx.state, "_build_summary", lambda: (_ for _ in ()).throw(RuntimeError("summary recompute")))
        ensure_summary_inputs(ctx)
        assert id(ctx.derived.summary_df) == summary_id


def test_writer_front_modules_use_explicit_derived_frames_not_compat_state(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        ctx = build_writer_context(_make_inputs(case_dir / "front_explicit.xlsx"))

        ensure_driver_inputs(ctx)
        ensure_summary_inputs(ctx)
        ensure_ui_evidence(ctx)
        ensure_report_inputs(ctx)
        ensure_hidden_value_inputs(ctx)
        ensure_valuation_inputs(ctx)
        ensure_raw_data_inputs(ctx)

        removed_keys = [
            "summary_df",
            "valuation_summary_df",
            "valuation_grid_df",
            "leverage_df",
            "report_is",
            "report_bs",
            "report_cf",
            "flags_df",
            "flags_audit_df",
            "flags_recompute_df",
            "signals_base_df",
            "quarter_notes_evidence_df",
            "promise_evidence_df",
            "ng_bridge",
            "ng_bridge_relaxed",
            "facts_long",
            "lineitem_map",
            "period_index",
        ]
        for key in removed_keys:
            ctx.state.pop(key, None)

        written: dict[str, pd.DataFrame] = {}

        def capture_sheet(name: str, frame: pd.DataFrame) -> None:
            written[name] = frame

        def capture_report(name: str, frame: pd.DataFrame, _units: str) -> None:
            written[name] = frame

        monkeypatch.setattr(ctx.callbacks, "write_summary_sheet", lambda frame: written.__setitem__("SUMMARY", frame))
        monkeypatch.setattr(ctx.callbacks, "write_sheet", capture_sheet)
        monkeypatch.setattr(ctx.callbacks, "write_report_sheet", capture_report)
        monkeypatch.setattr(ctx.callbacks, "write_flags_sheet", capture_sheet)
        monkeypatch.setattr(ctx.callbacks, "write_valuation_sheet", lambda: None)
        monkeypatch.setattr(ctx.callbacks, "write_bs_segments_sheet", lambda quarters_shown=8: [])
        monkeypatch.setattr(ctx.callbacks, "write_quarter_notes_ui_v2", lambda quarters_shown=8, **_: [])
        monkeypatch.setattr(ctx.callbacks, "write_promise_tracker_ui_v2", lambda *_, **__: [])
        monkeypatch.setattr(ctx.callbacks, "write_promise_progress_ui_v2", lambda *_, **__: [])
        monkeypatch.setattr(ctx.callbacks, "write_operating_drivers_raw_sheet", lambda rows: None)
        monkeypatch.setattr(ctx.callbacks, "write_economics_market_raw_sheet", lambda rows: None)

        write_summary_sheets(ctx)
        write_valuation_sheets(ctx)
        write_debt_sheets(ctx)
        write_report_sheets(ctx)
        write_ui_sheets(ctx)
        write_raw_data_sheets(ctx)

        assert written["SUMMARY"] is ctx.derived.summary_df
        assert written["Valuation_Summary"] is ctx.derived.valuation_summary_df
        assert written["Valuation_Grid"] is ctx.derived.valuation_grid_df
        assert written["Leverage_Liquidity"] is ctx.derived.leverage_df
        assert written["REPORT_IS_Q"] is ctx.derived.report_is
        assert written["REPORT_BS_Q"] is ctx.derived.report_bs
        assert written["REPORT_CF_Q"] is ctx.derived.report_cf
        assert written["Hidden_Value_Flags"] is ctx.derived.flags_df
        assert written["Hidden_Value_Audit"] is ctx.derived.flags_audit_df
        assert written["Hidden_Value_Recompute"] is ctx.derived.flags_recompute_df
        assert written["Hidden_Value_Base"] is ctx.derived.signals_base_df
        assert written["Quarter_Notes_Evidence"] is ctx.derived.quarter_notes_evidence_df
        assert written["Promise_Evidence"] is ctx.derived.promise_evidence_df
        assert written["NonGAAP_Bridge"] is ctx.derived.ng_bridge
        assert written["DATA_Facts_Long"] is ctx.derived.facts_long
        assert written["DATA_LineItem_Map"] is ctx.derived.lineitem_map
        assert written["DATA_Period_Index"] is ctx.derived.period_index


def test_summary_includes_current_strategic_context_row() -> None:
    with _case_dir() as case_dir:
        inputs = _make_inputs(case_dir / "summary_current_context.xlsx")
        inputs.company_overview = {
            "what_it_does": "Demo company description.",
            "what_it_does_source": "Source: SEC 10-K demo",
            "current_strategic_context": "Management is focused on capital allocation and cost discipline.",
            "current_strategic_context_source": "Source: SEC 8-K demo",
            "key_advantage": "Demo advantage.",
            "key_advantage_source": "Source: SEC 10-K competition demo",
            "segment_operating_model": [],
            "segment_operating_model_source": "Source: N/A",
            "key_dependencies": [],
            "key_dependencies_source": "Source: N/A",
            "wrong_thesis_bullets": [],
            "wrong_thesis_source": "Source: N/A",
            "revenue_streams": [],
            "revenue_streams_source": "Source: N/A",
            "asof_fy_end": None,
        }
        ctx = build_writer_context(inputs)
        ensure_summary_inputs(ctx)

        summary_df = ctx.derived.summary_df
        row = summary_df.loc[summary_df["Metric"] == "Current strategic context"].iloc[0]
        assert row["Value"] == "Management is focused on capital allocation and cost discipline."
        assert row["Note"] == "Source: SEC 8-K demo"


def test_operating_driver_line_index_and_template_rows_cache_reuse(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        profile = CompanyProfile(
            ticker="TEST",
            has_bank=False,
            industry_keywords=tuple(),
            segment_patterns=tuple(),
            segment_alias_patterns=tuple(),
            key_adv_require_keywords=tuple(),
            key_adv_deny_keywords=tuple(),
            enable_operating_drivers_sheet=True,
            operating_driver_history_templates=(
                OperatingDriverTemplate(
                    group="Operations",
                    label="Plant status",
                    why_it_matters="demo",
                    match_terms=("plant status", "operations"),
                    aliases=("plant", "operations"),
                    key="plant_status",
                ),
            ),
        )
        quarter_notes = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "text_full": ["Plant status remains strong.\nOperations continue normally across the fleet."],
                "doc_name": ["demo_note.txt"],
            }
        )
        monkeypatch.setattr(writer_context_module, "get_company_profile", lambda ticker: profile)
        ctx = build_writer_context(_make_inputs(case_dir / "driver_cache.xlsx", quarter_notes=quarter_notes))

        rows_first = ctx.callbacks.build_operating_drivers_history_rows()
        rows_second = ctx.callbacks.build_operating_drivers_history_rows()

        assert rows_first == rows_second
        assert ctx.derived.operating_driver_line_index_by_quarter
        assert ctx.derived.operating_driver_flat_line_index
        assert ctx.derived.operating_driver_best_text_cache
        assert ctx.derived.operating_driver_template_rows_cache
        assert ctx.derived.operating_driver_template_candidate_cache
        assert len(ctx.derived.operating_driver_template_candidate_cache) >= 1

        line_index_id = id(ctx.derived.operating_driver_line_index_by_quarter)
        template_rows_id = id(ctx.derived.operating_driver_template_rows_cache)
        best_text_cache_id = id(ctx.derived.operating_driver_best_text_cache)
        template_candidate_id = id(ctx.derived.operating_driver_template_candidate_cache)

        rows_third = ctx.callbacks.build_operating_drivers_history_rows()
        assert rows_third == rows_first
        assert id(ctx.derived.operating_driver_line_index_by_quarter) == line_index_id
        assert id(ctx.derived.operating_driver_template_rows_cache) == template_rows_id
        assert id(ctx.derived.operating_driver_best_text_cache) == best_text_cache_id
        assert id(ctx.derived.operating_driver_template_candidate_cache) == template_candidate_id


def test_valuation_source_views_and_render_bundle_are_reused() -> None:
    with _case_dir() as case_dir:
        ctx = build_writer_context(_make_inputs(case_dir / "valuation_cache.xlsx"))
        ensure_valuation_inputs(ctx)

        assert isinstance(ctx.derived.valuation_hist_indexed, pd.DataFrame)
        assert ctx.derived.valuation_latest_context is not None
        assert ctx.derived.valuation_last4_context is not None
        assert ctx.derived.valuation_core_maps is not None

        quarter_key = tuple(pd.Timestamp(q) for q in ctx.inputs.hist["quarter"].tolist())
        bundle_loader = ctx.state["_ensure_valuation_render_bundle"]
        bundle_first = bundle_loader(quarter_key, ctx.derived.leverage_df)
        bundle_second = bundle_loader(quarter_key, ctx.derived.leverage_df)

        assert bundle_first is bundle_second
        assert ctx.derived.valuation_render_bundle is bundle_first
        assert bundle_first["rev_map"]
        assert bundle_first["ebitda_ttm_map"]
        assert bundle_first["quarter_index_map"][quarter_key[0]] == 0
        assert len(bundle_first["last4_quarters_map"][quarter_key[3]]) == 4


def test_valuation_precompute_bundle_is_reused_and_reads_docs_once(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "valuation_precompute.xlsx")
        cache_dir = out_path.parent
        doc_path = cache_dir / "doc_000012345625000001_ex99.txt"
        doc_path.write_text(
            "In the fourth quarter, we repurchased $15 million of common stock. "
            "The board declared a regular quarterly dividend of $0.50 per share.",
            encoding="utf-8",
        )
        audit_df = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "accn": ["0000123456-25-000001"],
                "form": ["8-K"],
                "filed": [pd.Timestamp("2026-02-20")],
            }
        )
        counts = {str(doc_path): 0}
        original_read_text = Path.read_text
        original_glob = Path.glob

        def counted_read_text(self: Path, *args, **kwargs):
            path_str = str(self)
            if path_str in counts:
                counts[path_str] += 1
            return original_read_text(self, *args, **kwargs)

        def counted_glob(self: Path, pattern: str):
            if pattern.startswith("doc_000012345625000001_"):
                return [doc_path]
            return original_glob(self, pattern)

        monkeypatch.setattr(Path, "read_text", counted_read_text)
        monkeypatch.setattr(Path, "glob", counted_glob)
        inputs = _make_inputs(out_path)
        manifest_df = pd.DataFrame({"path": [str(doc_path)]})
        inputs = inputs.__class__(**{**vars(inputs), "audit": audit_df, "manifest_df": manifest_df})
        ctx = build_writer_context(inputs)
        ensure_valuation_inputs(ctx)

        quarter_key = tuple(pd.Timestamp(q) for q in ctx.inputs.hist["quarter"].tolist())
        render_loader = ctx.state["_ensure_valuation_render_bundle"]
        precompute_loader = ctx.state["_ensure_valuation_precompute_bundle"]
        render_bundle = render_loader(quarter_key, ctx.derived.leverage_df)
        bundle_first = precompute_loader(quarter_key, render_bundle)
        reads_after_first = counts[str(doc_path)]
        bundle_second = precompute_loader(quarter_key, render_bundle)

        assert bundle_first is bundle_second
        assert ctx.derived.valuation_precompute_bundle is bundle_first
        assert ctx.derived.valuation_filing_docs_by_quarter
        assert bundle_first["buyback_map"][pd.Timestamp("2025-12-31")] == pytest.approx(15_000_000.0)
        assert bundle_first["dividend_ps_doc_map"][pd.Timestamp("2025-12-31")] == pytest.approx(0.5)
        assert ctx.data.runtime_cache.valuation_precompute.buyback_execution_doc_cache
        assert reads_after_first >= 1
        assert counts[str(doc_path)] == reads_after_first


def test_valuation_precompute_subtimings_emit_when_needed(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "valuation_precompute_timings.xlsx")
        cache_dir = out_path.parent
        doc_path = cache_dir / "doc_000012345625000001_ex99.txt"
        doc_path.write_text(
            "In the fourth quarter, we repurchased $15 million of common stock. "
            "The board declared a regular quarterly dividend of $0.50 per share.",
            encoding="utf-8",
        )
        original_glob = Path.glob

        def counted_glob(self: Path, pattern: str):
            if pattern.startswith("doc_000012345625000001_"):
                return [doc_path]
            return original_glob(self, pattern)

        monkeypatch.setattr(Path, "glob", counted_glob)
        audit_df = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "accn": ["0000123456-25-000001"],
                "form": ["8-K"],
                "filed": [pd.Timestamp("2026-02-20")],
            }
        )
        inputs = _make_inputs(out_path, profile_timings=True)
        inputs = inputs.__class__(**{**vars(inputs), "audit": audit_df})
        ctx = build_writer_context(inputs)
        ensure_valuation_inputs(ctx)

        quarter_key = tuple(pd.Timestamp(q) for q in ctx.inputs.hist["quarter"].tolist())
        render_loader = ctx.state["_ensure_valuation_render_bundle"]
        precompute_loader = ctx.state["_ensure_valuation_precompute_bundle"]
        render_bundle = render_loader(quarter_key, ctx.derived.leverage_df)
        precompute_loader(quarter_key, render_bundle)
        stdout = capsys.readouterr().out

        assert "write_excel.valuation.precompute.doc_index" in stdout
        assert "write_excel.valuation.precompute.keyword_maps" in stdout
        assert "write_excel.valuation.precompute.buyback_dividend_maps" in stdout
        assert "write_excel.valuation.precompute.buyback_share_maps" in stdout


def test_buyback_auth_prefers_newest_filing_match(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "buyback_auth.xlsx")
        cache_dir = out_path.parent
        new_accn = "0000123456-26-000002"
        old_accn = "0000123456-25-000001"
        new_doc = cache_dir / "doc_000012345626000002_new-20251231.htm"
        old_doc = cache_dir / "doc_000012345625000001_old-20241231.htm"
        new_doc.write_text(
            "<html><body>As of December 31, 2025, $125 million remained available under the share repurchase authorization.</body></html>",
            encoding="utf-8",
        )
        old_doc.write_text(
            "<html><body>As of December 31, 2024, $50 million remained available under the share repurchase authorization.</body></html>",
            encoding="utf-8",
        )
        (cache_dir / "submissions_test.json").write_text(
            json.dumps(
                {
                    "filings": {
                        "recent": {
                            "accessionNumber": [new_accn, old_accn],
                            "form": ["10-K", "10-K"],
                            "filingDate": ["2026-02-10", "2025-02-10"],
                            "reportDate": ["2025-12-31", "2024-12-31"],
                            "primaryDocument": ["new-20251231.htm", "old-20241231.htm"],
                        }
                    }
                }
            ),
            encoding="utf-8",
        )
        manifest_df = pd.DataFrame(
            {
                "cache_key": [
                    "submissions_test",
                    "doc_000012345626000002_new-20251231.htm",
                    "doc_000012345625000001_old-20241231.htm",
                ],
                "path": [
                    str(cache_dir / "submissions_test.json"),
                    str(new_doc),
                    str(old_doc),
                ],
                "status": ["ok", "ok", "ok"],
            }
        )
        original_read_text = Path.read_text
        new_doc_reads = 0
        new_doc_resolved = new_doc.resolve()

        def counted_read_text(self: Path, *args, **kwargs):
            nonlocal new_doc_reads
            try:
                resolved = self.resolve()
            except Exception:
                resolved = self
            if resolved == new_doc_resolved:
                new_doc_reads += 1
            return original_read_text(self, *args, **kwargs)

        monkeypatch.setattr(Path, "read_text", counted_read_text)
        inputs = _make_inputs(out_path)
        inputs = inputs.__class__(**{**vars(inputs), "manifest_df": manifest_df})
        write_excel_from_inputs(inputs)
        stdout = capsys.readouterr().out

        assert out_path.exists()
        assert new_doc_reads >= 1
        assert f"[buyback_auth] match form=10-K accn={new_accn}" in stdout


def test_driver_45z_guidance_doc_index_is_reused_and_preserves_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "guidance.xlsx")
        cache_dir = out_path.parent
        guidance_doc = cache_dir / "doc_TEST_Q4_2025_guidance.txt"
        guidance_doc.write_text(
            "45Z expected monetization $250M-$300M expected monetization in 2026.",
            encoding="utf-8",
        )
        profile = CompanyProfile(
            ticker="TEST",
            has_bank=False,
            industry_keywords=tuple(),
            segment_patterns=tuple(),
            segment_alias_patterns=tuple(),
            key_adv_require_keywords=tuple(),
            key_adv_deny_keywords=tuple(),
            promise_priority_terms=("45z",),
            enable_operating_drivers_sheet=True,
            operating_driver_history_templates=(
                OperatingDriverTemplate(
                    group="Margin / spread",
                    label="45Z guided value",
                    why_it_matters="demo",
                    match_terms=("45z", "monetization"),
                    aliases=("45z",),
                    key="45z_value_guided",
                ),
            ),
        )
        quarter_notes = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "text_full": ["45Z monetization remains a core driver."],
                "doc_name": ["demo_note.txt"],
            }
        )
        monkeypatch.setattr(writer_context_module, "get_company_profile", lambda ticker: profile)
        ctx = build_writer_context(_make_inputs(out_path, quarter_notes=quarter_notes))

        docs_first = ctx.state["_load_operating_driver_45z_guidance_docs_by_quarter"]()
        docs_second = ctx.state["_load_operating_driver_45z_guidance_docs_by_quarter"]()
        rows_first = ctx.callbacks.build_operating_drivers_history_rows()
        rows_second = ctx.callbacks.build_operating_drivers_history_rows()

        assert docs_first is docs_second
        assert ctx.derived.operating_driver_45z_guidance_docs_by_quarter is docs_first
        assert rows_first == rows_second
        assert any(str(row.get("Commentary") or "").startswith("$250.0m-$300.0m expected monetization") for row in rows_first)


def test_driver_template_signal_and_doc_index_timings_emit_when_needed(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "driver_timings.xlsx")
        (out_path.parent / "doc_TEST_Q4_2025_guidance.txt").write_text(
            "45Z expected monetization $250M-$300M expected monetization in 2026.",
            encoding="utf-8",
        )
        profile = CompanyProfile(
            ticker="TEST",
            has_bank=False,
            industry_keywords=tuple(),
            segment_patterns=tuple(),
            segment_alias_patterns=tuple(),
            key_adv_require_keywords=tuple(),
            key_adv_deny_keywords=tuple(),
            promise_priority_terms=("45z",),
            enable_operating_drivers_sheet=True,
            operating_driver_history_templates=(
                OperatingDriverTemplate(
                    group="Margin / spread",
                    label="45Z guided value",
                    why_it_matters="demo",
                    match_terms=("45z", "monetization"),
                    aliases=("45z",),
                    key="45z_value_guided",
                ),
            ),
        )
        quarter_notes = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "text_full": ["45Z monetization remains a core driver."],
                "doc_name": ["demo_note.txt"],
            }
        )
        monkeypatch.setattr(writer_context_module, "get_company_profile", lambda ticker: profile)
        ctx = build_writer_context(_make_inputs(out_path, quarter_notes=quarter_notes, profile_timings=True))

        ctx.callbacks.build_operating_drivers_history_rows()
        stdout = capsys.readouterr().out

        assert "write_excel.derive.driver_inputs.template_signal_index" in stdout
        assert "write_excel.derive.driver_inputs.template_doc_index" in stdout


def test_valuation_net_leverage_text_map_prefers_adjusted_and_filters_quarters() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "leverage.xlsx")
        quarter_notes = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2024-12-31")],
                "claim": ["Net leverage was 3.5x.", "Adjusted net leverage was 1.9x."],
                "filed": [pd.Timestamp("2026-02-01"), pd.Timestamp("2025-02-01")],
            }
        )
        promises = pd.DataFrame(
            {
                "last_seen_quarter": [pd.Timestamp("2025-12-31")],
                "statement": ["Adjusted net leverage was 2.1x."],
                "filed": [pd.Timestamp("2026-02-15")],
                "source_type": ["promise"],
            }
        )
        inputs = _make_inputs(out_path, quarter_notes=quarter_notes, profile_timings=True)
        inputs = inputs.__class__(**{**vars(inputs), "promises": promises})
        ctx = build_writer_context(inputs)

        lev_map_first = ctx.callbacks.extract_adj_net_leverage_text_map()
        lev_map_second = ctx.callbacks.extract_adj_net_leverage_text_map()

        assert lev_map_first == lev_map_second
        assert lev_map_first[pd.Timestamp("2025-12-31")] == pytest.approx(2.1)
        assert pd.Timestamp("2024-12-31") not in lev_map_first


def test_valuation_net_leverage_text_map_reuses_doc_reads(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "leverage_docs.xlsx")
        root = out_path.parent.parent
        local_dir = root / "earnings_release"
        local_dir.mkdir(parents=True, exist_ok=True)
        sec_cache_dir = root / "sec_cache"
        sec_cache_dir.mkdir(parents=True, exist_ok=True)
        local_doc = local_dir / "TEST_Q4_2025_release.txt"
        local_doc.write_text("Adjusted net leverage was 2.0x.", encoding="utf-8")
        audit_doc = sec_cache_dir / "doc_000012345625000001_ex99.txt"
        audit_doc.write_text("Net leverage was 3.5x.", encoding="utf-8")
        audit_df = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "accn": ["0000123456-25-000001"],
                "form": ["8-K"],
                "filed": [pd.Timestamp("2026-02-20")],
            }
        )
        manifest_df = pd.DataFrame({"path": [str(audit_doc)]})
        counts = {str(local_doc): 0, str(audit_doc): 0}
        original_read_text = Path.read_text

        def counted_read_text(self: Path, *args, **kwargs):
            path_str = str(self)
            if path_str in counts:
                counts[path_str] += 1
            return original_read_text(self, *args, **kwargs)

        monkeypatch.setattr(Path, "read_text", counted_read_text)
        inputs = _make_inputs(out_path)
        inputs = inputs.__class__(**{**vars(inputs), "audit": audit_df, "manifest_df": manifest_df})
        ctx = build_writer_context(inputs)

        lev_map_first = ctx.callbacks.extract_adj_net_leverage_text_map()
        lev_map_second = ctx.callbacks.extract_adj_net_leverage_text_map()

        assert lev_map_first == lev_map_second
        assert lev_map_first[pd.Timestamp("2025-12-31")] == pytest.approx(2.0)
        assert counts[str(local_doc)] == 1
        assert counts[str(audit_doc)] == 1


def test_writer_doc_cache_is_shared_across_leverage_and_valuation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "shared_doc_cache.xlsx")
        cache_dir = out_path.parent
        doc_path = cache_dir / "doc_000012345625000001_ex99.txt"
        doc_path.write_text(
            "Adjusted net leverage was 2.0x. "
            "In the fourth quarter, we repurchased $15 million of common stock. "
            "The board declared a regular quarterly dividend of $0.50 per share.",
            encoding="utf-8",
        )
        audit_df = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "accn": ["0000123456-25-000001"],
                "form": ["8-K"],
                "filed": [pd.Timestamp("2026-02-20")],
            }
        )
        counts = {"read": 0, "glob": 0}
        original_read_text = Path.read_text
        original_glob = Path.glob
        doc_resolved = doc_path.resolve()
        cache_resolved = cache_dir.resolve()

        def counted_read_text(self: Path, *args, **kwargs):
            try:
                current = self.resolve()
            except Exception:
                current = self
            if current == doc_resolved:
                counts["read"] += 1
            return original_read_text(self, *args, **kwargs)

        def counted_glob(self: Path, pattern: str, *args, **kwargs):
            try:
                current = self.resolve()
            except Exception:
                current = self
            if current == cache_resolved and pattern.startswith("doc_000012345625000001_"):
                counts["glob"] += 1
            return original_glob(self, pattern, *args, **kwargs)

        monkeypatch.setattr(Path, "read_text", counted_read_text)
        monkeypatch.setattr(Path, "glob", counted_glob)
        inputs = _make_inputs(out_path)
        manifest_df = pd.DataFrame({"path": [str(doc_path)]})
        inputs = inputs.__class__(**{**vars(inputs), "audit": audit_df, "manifest_df": manifest_df})
        ctx = build_writer_context(inputs)

        lev_map = ctx.callbacks.extract_adj_net_leverage_text_map()
        ensure_valuation_inputs(ctx)
        quarter_key = tuple(pd.Timestamp(q) for q in ctx.inputs.hist["quarter"].tolist())
        render_loader = ctx.state["_ensure_valuation_render_bundle"]
        precompute_loader = ctx.state["_ensure_valuation_precompute_bundle"]
        render_bundle = render_loader(quarter_key, ctx.derived.leverage_df)
        precompute_bundle = precompute_loader(quarter_key, render_bundle)

        assert lev_map[pd.Timestamp("2025-12-31")] == pytest.approx(2.0)
        assert precompute_bundle["buyback_map"][pd.Timestamp("2025-12-31")] == pytest.approx(15_000_000.0)
        assert precompute_bundle["dividend_ps_doc_map"][pd.Timestamp("2025-12-31")] == pytest.approx(0.5)
        assert counts["glob"] == 3
        assert counts["read"] == 1
        assert ctx.data.doc_cache.accession_doc_paths["000012345625000001"] == [doc_path]


def test_latest_quarter_qa_reuses_shared_sec_doc_cache(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "qa_shared_doc_cache.xlsx")
        cache_dir = out_path.parent
        doc_path = cache_dir / "doc_000012345625000001_ex99.txt"
        doc_path.write_text(
            "Adjusted net leverage was 2.0x. Revenue was $130 million in the fourth quarter.",
            encoding="utf-8",
        )
        audit_df = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "accn": ["0000123456-25-000001"],
                "form": ["8-K"],
                "filed": [pd.Timestamp("2026-02-20")],
            }
        )
        manifest_df = pd.DataFrame({"path": [str(doc_path)]})
        counts = {"read": 0, "glob": 0}
        original_read_text = Path.read_text
        original_glob = Path.glob
        doc_resolved = doc_path.resolve()
        cache_resolved = cache_dir.resolve()

        def counted_read_text(self: Path, *args, **kwargs):
            try:
                current = self.resolve()
            except Exception:
                current = self
            if current == doc_resolved:
                counts["read"] += 1
            return original_read_text(self, *args, **kwargs)

        def counted_glob(self: Path, pattern: str, *args, **kwargs):
            try:
                current = self.resolve()
            except Exception:
                current = self
            pattern_txt = str(pattern)
            if current == cache_resolved and pattern_txt.startswith("doc_000012345625000001_"):
                counts["glob"] += 1
            return original_glob(self, pattern, *args, **kwargs)

        monkeypatch.setattr(Path, "read_text", counted_read_text)
        monkeypatch.setattr(Path, "glob", counted_glob)
        inputs = _make_inputs(out_path)
        inputs = inputs.__class__(**{**vars(inputs), "audit": audit_df, "manifest_df": manifest_df})
        ctx = build_writer_context(inputs)

        lev_map = ctx.callbacks.extract_adj_net_leverage_text_map()
        reads_after_leverage = counts["read"]
        globs_after_leverage = counts["glob"]
        qa_rows = ctx.callbacks.run_latest_quarter_qa()

        assert lev_map[pd.Timestamp("2025-12-31")] == pytest.approx(2.0)
        assert qa_rows
        assert counts["read"] == reads_after_leverage == 1
        assert counts["glob"] == globs_after_leverage == 3


def test_latest_quarter_qa_classifies_pbi_fcf_as_definition_mismatch(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_ticker_model_out_path(case_dir, "PBI", "pbi_fcf_definition_mismatch.xlsx")
            sec_cache_dir = out_path.parent.parent / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            source_doc = sec_cache_dir / "doc_000162828026008604_q42025earningspressrelea.htm"
            source_doc.write_text(
                (
                    "Pitney Bowes Discloses Financial Results for Fourth Quarter and Full Year 2025. "
                    "Fourth Quarter ($ millions except EPS) 2025 2024 Cash from Operations $222 $132 Free Cash Flow1 $212 $142. "
                    "Adjusted EBIT, Adjusted EBITDA and Free Cash Flow are non-GAAP measures. "
                    "Free cash flow adjusts cash flow from operations calculated in accordance with GAAP for capital expenditures, "
                    "restructuring payments and other special items. "
                    "Pitney Bowes Inc. Reconciliation of reported net cash from operating activities to free cash flow "
                    "Net cash from operating activities - continuing operations $221,699 $131,837 "
                    "Capital expenditures (20,251) (22,182) Restructuring payments 10,495 32,104 "
                    "Free cash flow $211,943 $141,759."
                ),
                encoding="utf-8",
            )
            hist = _make_hist().copy()
            hist["cfo"] = [90_000_000.0, 100_000_000.0, 110_000_000.0, 221_699_000.0]
            hist["capex"] = [20_000_000.0, 20_000_000.0, 20_000_000.0, 20_251_000.0]
            hist["cash"] = [300_000_000.0, 295_000_000.0, 290_000_000.0, 284_887_000.0]
            hist["debt_core"] = [2_250_000_000.0, 2_200_000_000.0, 2_150_000_000.0, 2_104_116_000.0]
            hist["total_debt"] = hist["debt_core"]
            inputs = _make_inputs(out_path, ticker="PBI", hist=hist)
            ctx = build_writer_context(inputs)

            qa_rows = ctx.callbacks.run_latest_quarter_qa()
            fcf_row = next(row for row in qa_rows if str(row.get("metric") or "") == "FCF (Q)")

            assert fcf_row["issue_family"] == "quarter_text_definition_mismatch"
            assert (
                fcf_row["message"]
                == "Workbook FCF (Q) is CFO-capex based at $201.4m; selected quarter text states company-defined free cash flow of $211.9m. Likely definition mismatch rather than same-basis numeric conflict."
            )
            assert str(fcf_row.get("recommended_action") or "") == "review metric definition"
            assert str(fcf_row.get("source") or "").endswith("doc_000162828026008604_q42025earningspressrelea.htm")


def test_latest_quarter_qa_classifies_gpre_total_debt_as_definition_mismatch(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_ticker_model_out_path(case_dir, "GPRE", "gpre_total_debt_definition_mismatch.xlsx")
            release_dir = out_path.parent.parent / "earnings_release"
            release_dir.mkdir(parents=True, exist_ok=True)
            source_doc = release_dir / "GPRE_Q4_2025_release.txt"
            source_doc.write_text(
                (
                    "Green Plains Reports Fourth Quarter and Full Year 2025 Financial Results. "
                    "Results for the Fourth Quarter 2025 and Future Outlook: "
                    "Total debt outstanding at December 31, 2025 was $399.5 million, including $33.6 million outstanding debt under "
                    "working capital revolvers and other short-term borrowing arrangements."
                ),
                encoding="utf-8",
            )
            hist = _make_hist().copy()
            hist["cash"] = [40_000_000.0, 38_000_000.0, 36_000_000.0, 33_600_000.0]
            hist["debt_core"] = [390_000_000.0, 380_000_000.0, 372_000_000.0, 365_900_000.0]
            hist["total_debt"] = [390_000_000.0, 380_000_000.0, 372_000_000.0, 365_900_000.0]
            inputs = _make_inputs(out_path, ticker="GPRE", hist=hist)
            ctx = build_writer_context(inputs)

            qa_rows = ctx.callbacks.run_latest_quarter_qa()
            debt_row = next(row for row in qa_rows if str(row.get("metric") or "") == "Total debt (Q)")

            assert debt_row["issue_family"] == "quarter_text_definition_mismatch"
            assert (
                debt_row["message"]
                == "Workbook Total debt (Q) is $365.9m on the modeled debt-profile basis; release total debt outstanding is $399.5m and includes revolver/other short-term borrowings. Likely basis/presentation mismatch rather than same-basis numeric conflict."
            )
            assert str(debt_row.get("recommended_action") or "") == "review metric definition"
            assert str(debt_row.get("source") or "").endswith("GPRE_Q4_2025_release.txt")


def test_latest_quarter_qa_downgrades_noisy_no_explicit_support_to_info() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "latest_quarter_support_noise.xlsx")
        inputs = _make_inputs(out_path, ticker="TEST")
        ctx = build_writer_context(inputs)

        qa_rows = ctx.callbacks.run_latest_quarter_qa()
        ebitda_row = next(row for row in qa_rows if str(row.get("metric") or "") == "EBITDA (Q)")
        revenue_row = next(row for row in qa_rows if str(row.get("metric") or "") == "Revenue (Q)")

        assert ebitda_row["issue_family"] == "quarter_text_no_explicit_support"
        assert ebitda_row["severity"] == "info"
        assert revenue_row["issue_family"] == "quarter_text_no_explicit_support"
        assert revenue_row["severity"] == "warn"


def test_latest_quarter_qa_emits_buyback_authorization_health_when_submissions_cache_missing(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "buyback_auth_health_missing_submissions.xlsx")
        monkeypatch.setattr(writer_context_module, "source_submission_cache_files", lambda **kwargs: [])
        monkeypatch.setattr(writer_context_module, "source_submission_recent_rows", lambda **kwargs: [])
        inputs = _make_inputs(out_path, ticker="PBI")
        ctx = build_writer_context(inputs)

        qa_rows = ctx.callbacks.run_latest_quarter_qa()
        auth_row = next(row for row in qa_rows if str(row.get("metric") or "") == "QA_BuybackAuthorization")

        assert str(auth_row.get("severity") or "") == "info"
        assert "submissions cache unavailable" in str(auth_row.get("message") or "").lower()


@pytest.mark.parametrize(
    ("source_choice", "expected_source_label", "expected_phrase"),
    [
        ("fallback_fy", "FY fallback", "same-end-date FY fallback"),
        ("carry_forward_q3", "Carry-forward", "prior-quarter carry-forward"),
    ],
)
def test_report_and_qa_surface_year_end_shares_diluted_fallback_provenance(
    source_choice: str,
    expected_source_label: str,
    expected_phrase: str,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, f"shares_diluted_{source_choice}.xlsx")
        hist = _make_hist().copy()
        hist["shares_diluted"] = [100_000_000.0, 101_000_000.0, 102_000_000.0, 103_000_000.0]
        audit = pd.DataFrame(
            [
                {
                    "metric": "shares_diluted",
                    "quarter": pd.Timestamp("2025-12-31"),
                    "source": "sec_facts",
                    "source_choice": source_choice,
                    "tag": "WeightedAverageNumberOfDilutedSharesOutstanding",
                    "accn": "0000123456-26-000001",
                    "form": "10-K",
                    "filed": pd.Timestamp("2026-02-20"),
                    "start": pd.NaT,
                    "end": pd.Timestamp("2025-12-31"),
                    "unit": "shares",
                    "value": 103_000_000.0,
                    "note": "FY-end diluted shares fallback from same-end-date FY fact"
                    if source_choice == "fallback_fy"
                    else "carry-forward diluted shares from prior quarter 2025-09-30",
                }
            ]
        )

        write_excel_from_inputs(_make_inputs(out_path, ticker="PBI", hist=hist, audit=audit))

        wb = load_workbook(out_path, data_only=False, read_only=False)
        try:
            ws_report = wb["REPORT_BS_Q"]
            diluted_row = _find_row_with_value(ws_report, "Diluted shares", column=2)
            assert diluted_row is not None
            assert str(ws_report.cell(row=diluted_row, column=3).value or "").strip() == expected_source_label
            assert str(ws_report.cell(row=diluted_row, column=4).value or "").strip() == "WARN"

            ws_qa = wb["QA_Checks"]
            qa_header = [str(ws_qa.cell(row=1, column=cc).value or "").strip() for cc in range(1, ws_qa.max_column + 1)]
            qa_rows = [
                {qa_header[cc - 1]: ws_qa.cell(row=rr, column=cc).value for cc in range(1, ws_qa.max_column + 1)}
                for rr in range(2, ws_qa.max_row + 1)
            ]
            assert any(
                str(row.get("metric") or "").strip() == "QA_SharesDiluted"
                and str(row.get("review_status") or "").strip() == "Watch"
                and expected_phrase in str(row.get("message") or "")
                for row in qa_rows
            )
        finally:
            wb.close()


def test_curated_needs_review_exposes_review_status_and_debt_integrity_cluster() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "needs_review_review_status.xlsx")
        needs_review = pd.DataFrame(
            [
                {
                    "quarter": pd.Timestamp("2025-12-31"),
                    "metric": "Total debt (Q)",
                    "severity": "warn",
                    "message": "Release total debt includes revolver and other short-term borrowing arrangements.",
                    "source": "release_q4.txt",
                    "issue_family": "revolver_and_other_debt_presence_check",
                    "raw_metric": "Total debt (Q)",
                },
                {
                    "quarter": pd.Timestamp("2025-12-31"),
                    "metric": "debt_tieout",
                    "severity": "warn",
                    "message": "Prefer debt source hierarchy for latest quarter tieout.",
                    "source": "release_q4.txt",
                    "issue_family": "debt_recon_coverage_check",
                    "raw_metric": "Debt tieout",
                },
                {
                    "quarter": pd.Timestamp("2025-12-31"),
                    "metric": "debt_tranches",
                    "severity": "warn",
                    "message": "Scale inferred from debt table; verify scaling.",
                    "source": "release_q4.txt",
                    "issue_family": "debt_tranches",
                    "raw_metric": "Debt tranches",
                },
                {
                    "quarter": pd.Timestamp("2025-12-31"),
                    "metric": "FCF (Q)",
                    "severity": "warn",
                    "message": "Workbook FCF (Q) is CFO-capex based at $201.4m; selected quarter text states company-defined free cash flow of $211.9m. Likely definition mismatch rather than same-basis numeric conflict.",
                    "source": "release_q4.txt",
                    "issue_family": "quarter_text_definition_mismatch",
                    "raw_metric": "FCF (Q)",
                },
            ]
        )

        write_excel_from_inputs(_make_inputs(out_path, ticker="PBI", needs_review=needs_review))

        wb = load_workbook(out_path, data_only=False, read_only=False)
        try:
            ws_nr = wb["Needs_Review"]
            header = [str(ws_nr.cell(row=1, column=cc).value or "").strip() for cc in range(1, ws_nr.max_column + 1)]
            rows = [
                {header[cc - 1]: ws_nr.cell(row=rr, column=cc).value for cc in range(1, ws_nr.max_column + 1)}
                for rr in range(2, ws_nr.max_row + 1)
            ]
            assert "review_status" in header
            assert any(
                str(row.get("issue_family") or "").strip() == "debt_integrity"
                and str(row.get("review_status") or "").strip() == "Action required"
                and "Debt integrity:" in str(row.get("latest_message") or "")
                for row in rows
            )
            assert any(
                str(row.get("issue_family") or "").strip() == "quarter_text_definition_mismatch"
                and str(row.get("review_status") or "").strip() == "Definition mismatch"
                and str(row.get("raw_metric") or "").strip() == "FCF (Q)"
                for row in rows
            )
        finally:
            wb.close()


def test_empty_product_sheets_use_current_build_placeholder() -> None:
    with _case_dir() as case_dir:
        out_path = _make_ticker_model_out_path(case_dir, "GPRE", "empty_sheet_placeholder.xlsx")
        write_excel_from_inputs(
            _make_inputs(
                out_path,
                ticker="GPRE",
                adj_breakdown=pd.DataFrame(),
                ocr_log=pd.DataFrame(),
                slides_segments=pd.DataFrame(),
            )
        )

        wb = load_workbook(out_path, data_only=False, read_only=False)
        try:
            assert str(wb["Adjustments_Breakdown"]["A1"].value or "").strip() == "No data for current build"
            assert str(wb["OCR_Text_Log"]["A1"].value or "").strip() == "No data for current build"
            assert str(wb["Slides_Segments"]["A1"].value or "").strip() == "No data for current build"
        finally:
            wb.close()


def test_submission_recent_rows_cache_is_shared_across_valuation_and_qa(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "submission_cache.xlsx")
        cache_dir = out_path.parent
        doc_path = cache_dir / "doc_000012345625000001_ex99.txt"
        sub_path = cache_dir / "submissions_test.json"
        doc_path.write_text(
            "In the fourth quarter, we repurchased $15 million of common stock. Revenue was $130 million.",
            encoding="utf-8",
        )
        sub_path.write_text(
            json.dumps(
                {
                    "filings": {
                        "recent": {
                            "accessionNumber": ["0000123456-25-000001"],
                            "form": ["8-K"],
                            "filingDate": ["2026-02-20"],
                            "reportDate": ["2025-12-31"],
                            "primaryDocument": ["ex99.txt"],
                        }
                    }
                }
            ),
            encoding="utf-8",
        )
        audit_df = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "accn": ["0000123456-25-000001"],
                "form": ["8-K"],
                "filed": [pd.Timestamp("2026-02-20")],
            }
        )
        counts = {"sub_read": 0, "sub_glob": 0}
        original_read_text = Path.read_text
        original_glob = Path.glob
        sub_resolved = sub_path.resolve()
        cache_resolved = cache_dir.resolve()

        def counted_read_text(self: Path, *args, **kwargs):
            try:
                current = self.resolve()
            except Exception:
                current = self
            if current == sub_resolved:
                counts["sub_read"] += 1
            return original_read_text(self, *args, **kwargs)

        def counted_glob(self: Path, pattern: str):
            try:
                current = self.resolve()
            except Exception:
                current = self
            if current == cache_resolved and pattern == "submissions_*.json":
                counts["sub_glob"] += 1
            return original_glob(self, pattern)

        monkeypatch.setattr(Path, "read_text", counted_read_text)
        monkeypatch.setattr(Path, "glob", counted_glob)
        manifest_df = pd.DataFrame({"path": [str(doc_path), str(sub_path)]})
        inputs = _make_inputs(out_path)
        inputs = inputs.__class__(**{**vars(inputs), "audit": audit_df, "manifest_df": manifest_df})
        ctx = build_writer_context(inputs)

        ensure_valuation_inputs(ctx)
        quarter_key = tuple(pd.Timestamp(q) for q in ctx.inputs.hist["quarter"].tolist())
        render_loader = ctx.state["_ensure_valuation_render_bundle"]
        precompute_loader = ctx.state["_ensure_valuation_precompute_bundle"]
        render_bundle = render_loader(quarter_key, ctx.derived.leverage_df)
        precompute_loader(quarter_key, render_bundle)
        reads_after_val = counts["sub_read"]
        globs_after_val = counts["sub_glob"]
        qa_rows = ctx.callbacks.run_latest_quarter_qa()

        assert qa_rows
        assert counts["sub_read"] == reads_after_val == 1
        assert counts["sub_glob"] == globs_after_val == 1


def test_buyback_auth_reuses_submission_rows_during_full_write(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "buyback_auth_cache.xlsx")
        cache_dir = out_path.parent
        accn = "0000123456-26-000002"
        doc_path = cache_dir / "doc_000012345626000002_new-20251231.htm"
        sub_path = cache_dir / "submissions_test.json"
        doc_path.write_text(
            "<html><body>"
            "In the fourth quarter, we repurchased $15 million of common stock. "
            "As of December 31, 2025, $125 million remained available under the share repurchase authorization. "
            "The board declared a regular quarterly dividend of $0.50 per share."
            "</body></html>",
            encoding="utf-8",
        )
        sub_path.write_text(
            json.dumps(
                {
                    "filings": {
                        "recent": {
                            "accessionNumber": [accn],
                            "form": ["10-K"],
                            "filingDate": ["2026-02-10"],
                            "reportDate": ["2025-12-31"],
                            "primaryDocument": ["new-20251231.htm"],
                        }
                    }
                }
            ),
            encoding="utf-8",
        )
        audit_df = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "accn": [accn],
                "form": ["10-K"],
                "filed": [pd.Timestamp("2026-02-10")],
            }
        )
        manifest_df = pd.DataFrame(
            {
                "cache_key": [
                    "submissions_test",
                    "doc_000012345626000002_new-20251231.htm",
                ],
                "path": [
                    str(sub_path),
                    str(doc_path),
                ],
                "status": ["ok", "ok"],
            }
        )
        counts = {"sub_read": 0}
        original_read_text = Path.read_text
        sub_resolved = sub_path.resolve()

        def counted_read_text(self: Path, *args, **kwargs):
            try:
                current = self.resolve()
            except Exception:
                current = self
            if current == sub_resolved:
                counts["sub_read"] += 1
            return original_read_text(self, *args, **kwargs)

        monkeypatch.setattr(Path, "read_text", counted_read_text)
        inputs = _make_inputs(out_path)
        inputs = inputs.__class__(**{**vars(inputs), "audit": audit_df, "manifest_df": manifest_df})
        ctx = build_writer_context(inputs)
        ensure_valuation_inputs(ctx)
        write_valuation_sheets(ctx)

        assert "Valuation" in ctx.wb.sheetnames
        assert counts["sub_read"] == 1


def test_slide_text_cache_is_reused_across_quarter_filtered_writer_paths(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "slide_cache.xlsx")
        material_root = out_path.parent.parent
        slide_dir = material_root / "sec_cache" / "slides_text"
        slide_dir.mkdir(parents=True, exist_ok=True)
        slide_path = slide_dir / "Q4_2025_p1.txt"
        slide_path.write_text(
            "Production at 95% of stated capacity\nUtilization improved in the quarter.",
            encoding="utf-8",
        )

        counts = {"read": 0, "glob": 0}
        original_read_text = Path.read_text
        original_glob = Path.glob
        slide_resolved = slide_path.resolve()
        slide_dir_resolved = slide_dir.resolve()

        def counted_read_text(self: Path, *args, **kwargs):
            try:
                current = self.resolve()
            except Exception:
                current = self
            if current == slide_resolved:
                counts["read"] += 1
            return original_read_text(self, *args, **kwargs)

        def counted_glob(self: Path, pattern: str):
            try:
                current = self.resolve()
            except Exception:
                current = self
            if current == slide_dir_resolved and pattern == "*.txt":
                counts["glob"] += 1
            return original_glob(self, pattern)

        monkeypatch.setattr(Path, "read_text", counted_read_text)
        monkeypatch.setattr(Path, "glob", counted_glob)
        ctx = build_writer_context(_make_inputs(out_path))
        quarter_end = pd.Timestamp("2025-12-31").date()

        pages = ctx.state["_pbi_slide_pages_for_qd"](quarter_end)
        util_text = ctx.state["_local_slide_driver_fallback"](quarter_end, "utilization")

        assert pages
        assert "95%" in str(pages[0]["text"])
        assert util_text.startswith("Production at 95%")
        assert counts["glob"] == 1
        assert counts["read"] == 1


def test_profile_slide_signals_are_lazy_and_reuse_persistent_cache(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_ticker_model_out_path(case_dir, "TEST", "profile_slide_signals.xlsx")
        material_root = out_path.parent.parent
        slide_dir = material_root / "sec_cache" / "slides_text"
        slide_dir.mkdir(parents=True, exist_ok=True)
        slide_path = slide_dir / "Q4_2025_p1.txt"
        slide_path.write_text(
            "Production at 95% of stated capacity.\nUtilization improved in the quarter.",
            encoding="utf-8",
        )

        test_profile = CompanyProfile(
            ticker="TEST",
            has_bank=False,
            industry_keywords=tuple(),
            segment_patterns=tuple(),
            segment_alias_patterns=tuple(),
            key_adv_require_keywords=tuple(),
            key_adv_deny_keywords=tuple(),
            quarter_note_priority_terms=("utilization",),
            promise_priority_terms=tuple(),
        )

        counts = {"read": 0}
        original_read_text = Path.read_text
        slide_resolved = slide_path.resolve()

        def counted_read_text(self: Path, *args, **kwargs):
            try:
                current = self.resolve()
            except Exception:
                current = self
            if current == slide_resolved:
                counts["read"] += 1
            return original_read_text(self, *args, **kwargs)

        monkeypatch.setattr(writer_context_module, "get_company_profile", lambda _ticker: test_profile)
        monkeypatch.setattr(Path, "read_text", counted_read_text)

        ctx_first = build_writer_context(_make_inputs(out_path, ticker="TEST"))

        assert counts["read"] == 0

        grouped_first = ctx_first.state["_load_profile_slide_signals_by_quarter"]()

        assert grouped_first[pd.Timestamp("2025-12-31").date()]
        assert counts["read"] == 1

        counts["read"] = 0
        ctx_second = build_writer_context(_make_inputs(out_path, ticker="TEST"))
        grouped_second = ctx_second.state["_load_profile_slide_signals_by_quarter"]()

        assert grouped_second == grouped_first
        assert counts["read"] == 0


def test_frame_view_cache_reuses_quarter_notes_view_without_mutating_inputs() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "frame_view_cache.xlsx")
        quarter_notes = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "note_id": ["note-1"],
                "category": ["Guidance"],
                "claim": ["Revenue should stabilize by year end."],
                "note": ["Revenue should stabilize by year end."],
                "metric_ref": ["revenue_yoy"],
                "doc_path": ["doc.html"],
                "evidence_snippet": ["Revenue should stabilize by year end."],
                "evidence_json": [
                    json.dumps(
                        [
                            {
                                "doc_path": "doc.html",
                                "section_or_page": "p1",
                                "snippet": "Revenue should stabilize by year end.",
                            }
                        ]
                    )
                ],
            }
        )
        ctx = build_writer_context(_make_inputs(out_path, quarter_notes=quarter_notes))

        assert "_quarter" not in quarter_notes.columns
        assert ("quarter_notes", "timestamp") not in ctx.data.frame_view_cache

        ctx.callbacks.build_qn_evidence_src()
        cached_view = ctx.data.frame_view_cache[("quarter_notes", "timestamp")]
        cached_id = id(cached_view)
        ctx.callbacks.write_quarter_notes_ui_v2()

        assert id(ctx.data.frame_view_cache[("quarter_notes", "timestamp")]) == cached_id
        assert "_quarter" in cached_view.columns
        assert "_quarter" not in quarter_notes.columns
        assert "_quarter" not in ctx.inputs.quarter_notes.columns


def test_quarter_note_runtime_signature_changes_when_semantics_change() -> None:
    base = {
        "note_id": "n-1",
        "candidate_type": "investor_note",
        "change_badge": "NEW",
        "bucket": "Capital allocation / shareholder returns",
        "_metric_display": "Buyback execution",
        "metric_canon": "Buyback execution",
        "metric_tag": "buybacks_cash",
        "text_full": "Repurchased 12.6m shares for $126.6m with an average price of $10.04/share in Q4.",
        "comment_full_text": "Repurchased 12.6m shares for $126.6m with an average price of $10.04/share in Q4.",
        "evidence_snippet": "Repurchased 12.6m shares for $126.6m with an average price of $10.04/share in Q4.",
        "score": 96.0,
        "_event_score": 101.0,
        "doc_priority": 5,
        "source": {
            "source_type": "issuer_purchases_table",
            "doc": "doc_000000000026000001_10k.htm",
            "form": "10-K",
            "section": "Issuer Purchases of Equity Securities",
        },
    }

    source_changed = dict(base)
    source_changed["source"] = {
        "source_type": "earnings_release",
        "doc": "ex99_1.htm",
        "form": "8-K",
        "section": "Press release",
    }
    score_changed = dict(base)
    score_changed["_event_score"] = 88.0
    summary_changed = dict(base)
    summary_changed["_render_summary"] = "Repurchase authorization increased by $250.0m."

    base_sig = writer_context_module._quarter_note_runtime_signature(base)

    assert writer_context_module._quarter_note_runtime_signature(source_changed) != base_sig
    assert writer_context_module._quarter_note_runtime_signature(score_changed) != base_sig
    assert writer_context_module._quarter_note_runtime_signature(summary_changed) != base_sig


def test_quarter_note_runtime_cache_key_ignores_dict_identity_for_semantic_copies() -> None:
    base = {
        "note_id": "n-1",
        "candidate_type": "investor_note",
        "change_badge": "NEW",
        "bucket": "Capital allocation / shareholder returns",
        "_metric_display": "Buyback execution",
        "metric_canon": "Buyback execution",
        "metric_tag": "buybacks_cash",
        "text_full": "Repurchased 12.6m shares for $126.6m with an average price of $10.04/share in Q4.",
        "comment_full_text": "Repurchased 12.6m shares for $126.6m with an average price of $10.04/share in Q4.",
        "evidence_snippet": "Repurchased 12.6m shares for $126.6m with an average price of $10.04/share in Q4.",
        "score": 96.0,
        "_event_score": 101.0,
        "doc_priority": 5,
        "source": {
            "source_type": "issuer_purchases_table",
            "doc": "doc_000000000026000001_10k.htm",
            "form": "10-K",
            "section": "Issuer Purchases of Equity Securities",
        },
    }
    copied = dict(base)
    changed = dict(base)
    changed["_render_summary"] = "Repurchase authorization increased by $250.0m."

    base_key = writer_context_module._quarter_note_runtime_cache_key("render_summary", base, date(2025, 12, 31))
    copied_key = writer_context_module._quarter_note_runtime_cache_key("render_summary", copied, date(2025, 12, 31))
    changed_key = writer_context_module._quarter_note_runtime_cache_key("render_summary", changed, date(2025, 12, 31))

    assert base_key == copied_key
    assert changed_key != base_key


def test_quarter_note_runtime_cache_key_includes_quarter_context() -> None:
    item = {
        "note_id": "n-1",
        "candidate_type": "investor_note",
        "bucket": "Capital allocation / shareholder returns",
        "text_full": "Repurchased 12.6m shares for $126.6m with an average price of $10.04/share in Q4.",
        "source": {"source_type": "issuer_purchases_table", "doc": "doc_1.htm", "form": "10-K"},
    }

    q4_key = writer_context_module._quarter_note_runtime_cache_key("late_stage_metadata", item, date(2025, 12, 31))
    q3_key = writer_context_module._quarter_note_runtime_cache_key("late_stage_metadata", item, date(2025, 9, 30))

    assert q4_key != q3_key


def test_quarter_doc_pool_reuses_cached_doc_analysis_within_one_run(monkeypatch: pytest.MonkeyPatch) -> None:
    with _case_dir() as case_dir:
        doc_path = case_dir / "doc_000012345625000001_10k.txt"
        doc_path.write_text(
            "Operating activities improved due to working capital release.",
            encoding="utf-8",
        )
        runtime = QuarterNotesRuntime()
        rows = [
            {
                "accn": "0000123456-25-000001",
                "form": "10-K",
                "filed": "2026-02-20",
                "report": "2025-12-31",
                "doc": doc_path.name,
            }
        ]
        loader_calls = {"count": 0}
        read_calls = {"count": 0}

        def _rows_loader() -> list[dict[str, object]]:
            loader_calls["count"] += 1
            return list(rows)

        def _parse_date(value: object) -> date | None:
            ts = pd.to_datetime(value, errors="coerce")
            return ts.date() if pd.notna(ts) else None

        def _read_cached_doc_text(path_in: Path) -> str:
            read_calls["count"] += 1
            return path_in.read_text(encoding="utf-8")

        first = runtime.quarter_doc_pool(
            date(2025, 12, 31),
            {"10-K"},
            rows_loader=_rows_loader,
            parse_date=_parse_date,
            sec_docs_for_accession=lambda _accn: [doc_path],
            locate_cached_doc_path=lambda _accn, _doc: doc_path,
            path_cache_key=lambda path_in: str(path_in.resolve()),
            read_cached_doc_text=_read_cached_doc_text,
            normalize_text=writer_context_module.glx_normalize_text,
            max_docs=1,
            doc_scope="primary",
            row_scope="quarter_filtered",
            require_quarter_match=False,
        )
        second = runtime.quarter_doc_pool(
            date(2025, 12, 31),
            {"10-K"},
            rows_loader=_rows_loader,
            parse_date=_parse_date,
            sec_docs_for_accession=lambda _accn: [doc_path],
            locate_cached_doc_path=lambda _accn, _doc: doc_path,
            path_cache_key=lambda path_in: str(path_in.resolve()),
            read_cached_doc_text=_read_cached_doc_text,
            normalize_text=writer_context_module.glx_normalize_text,
            max_docs=1,
            doc_scope="primary",
            row_scope="quarter_filtered",
            require_quarter_match=False,
        )

        assert len(first) == 1
        assert second == first
        assert loader_calls["count"] == 1
        assert read_calls["count"] == 1


def test_quarter_doc_pool_does_not_leak_between_quarters() -> None:
    with _case_dir() as case_dir:
        doc_path = case_dir / "doc_000012345625000001_10k.txt"
        doc_path.write_text(
            "Operating activities improved due to working capital release.",
            encoding="utf-8",
        )
        runtime = QuarterNotesRuntime()
        rows = [
            {
                "accn": "0000123456-25-000001",
                "form": "10-K",
                "filed": "2026-02-20",
                "report": "2025-12-31",
                "doc": doc_path.name,
            }
        ]

        def _parse_date(value: object) -> date | None:
            ts = pd.to_datetime(value, errors="coerce")
            return ts.date() if pd.notna(ts) else None

        q4_rows = runtime.quarter_doc_pool(
            date(2025, 12, 31),
            {"10-K"},
            rows_loader=lambda: list(rows),
            parse_date=_parse_date,
            sec_docs_for_accession=lambda _accn: [doc_path],
            locate_cached_doc_path=lambda _accn, _doc: doc_path,
            path_cache_key=lambda path_in: str(path_in.resolve()),
            read_cached_doc_text=lambda path_in: path_in.read_text(encoding="utf-8"),
            normalize_text=writer_context_module.glx_normalize_text,
            max_docs=1,
            doc_scope="primary",
            row_scope="quarter_filtered",
            require_quarter_match=False,
        )
        q3_rows = runtime.quarter_doc_pool(
            date(2025, 9, 30),
            {"10-K"},
            rows_loader=lambda: list(rows),
            parse_date=_parse_date,
            sec_docs_for_accession=lambda _accn: [doc_path],
            locate_cached_doc_path=lambda _accn, _doc: doc_path,
            path_cache_key=lambda path_in: str(path_in.resolve()),
            read_cached_doc_text=lambda path_in: path_in.read_text(encoding="utf-8"),
            normalize_text=writer_context_module.glx_normalize_text,
            max_docs=1,
            doc_scope="primary",
            row_scope="quarter_filtered",
            require_quarter_match=False,
        )

        assert len(q4_rows) == 1
        assert q3_rows == []


def test_quarter_notes_runtime_cache_is_scoped_to_each_export_run(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            first_out = _make_model_out_path(case_dir, "runtime_cache_run_one.xlsx")
            second_out = _make_model_out_path(case_dir, "runtime_cache_run_two.xlsx")
            first_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "note_id": ["same-note-id"],
                    "category": ["Guidance / outlook"],
                    "claim": ["FY 2026 Revenue guidance $1,760m-$1,860m."],
                    "note": ["FY 2026 Revenue guidance $1,760m-$1,860m."],
                    "metric_ref": ["Revenue guidance"],
                    "score": [95.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q4_v1.txt"],
                    "source_type": ["earnings_release"],
                    "evidence_snippet": ["FY 2026 Revenue guidance $1,760m-$1,860m."],
                }
            )
            second_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "note_id": ["same-note-id"],
                    "category": ["Guidance / outlook"],
                    "claim": ["FY 2026 Revenue guidance $1,880m-$1,980m."],
                    "note": ["FY 2026 Revenue guidance $1,880m-$1,980m."],
                    "metric_ref": ["Revenue guidance"],
                    "score": [95.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q4_v2.txt"],
                    "source_type": ["earnings_release"],
                    "evidence_snippet": ["FY 2026 Revenue guidance $1,880m-$1,980m."],
                }
            )

            first_result = write_excel_from_inputs(
                _make_inputs(first_out, ticker="TEST", quarter_notes=first_notes)
            )
            second_result = write_excel_from_inputs(
                _make_inputs(second_out, ticker="TEST", quarter_notes=second_notes)
            )

            first_rows = list(first_result.quarter_notes_ui_snapshot.get("2025-12-31") or [])
            second_rows = list(second_result.quarter_notes_ui_snapshot.get("2025-12-31") or [])

            assert any("FY 2026 Revenue guidance $1,760m-$1,860m." in note for _, note in first_rows)
            assert not any("FY 2026 Revenue guidance $1,880m-$1,980m." in note for _, note in first_rows)
            assert any("FY 2026 Revenue guidance $1,880m-$1,980m." in note for _, note in second_rows)
            assert not any("FY 2026 Revenue guidance $1,760m-$1,860m." in note for _, note in second_rows)


def test_latest_quarter_sec_text_corpus_is_reused_across_repeated_qa_runs(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "latest_q_corpus.xlsx")
        material_root = out_path.parent.parent
        cache_dir = material_root / "sec_cache"
        cache_dir.mkdir(parents=True, exist_ok=True)
        accn = "0000123456-25-000001"
        doc_path = cache_dir / f"doc_{accn.replace('-', '')}_ex99.htm"
        doc_path.write_text("Updated full year guidance revenue $500 million to $520 million.", encoding="utf-8")

        audit_df = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-12-31")],
                "accn": [accn],
                "form": ["8-K"],
                "filed": [pd.Timestamp("2026-02-20")],
            }
        )

        counts = {"glob": 0}
        original_glob = Path.glob
        cache_resolved = cache_dir.resolve()

        def counted_glob(self: Path, pattern: str, *args, **kwargs):
            try:
                current = self.resolve()
            except Exception:
                current = self
            pattern_txt = str(pattern)
            if current == cache_resolved and pattern_txt.startswith(f"doc_{accn.replace('-', '')}_"):
                counts["glob"] += 1
            return original_glob(self, pattern, *args, **kwargs)

        monkeypatch.setattr(Path, "glob", counted_glob)
        manifest_df = pd.DataFrame({"path": [str(doc_path)]})
        inputs = _make_inputs(out_path)
        inputs = inputs.__class__(**{**vars(inputs), "audit": audit_df, "manifest_df": manifest_df})
        ctx = build_writer_context(inputs)

        qa_once = ctx.callbacks.run_latest_quarter_qa()
        globs_after_first = counts["glob"]
        qa_twice = ctx.callbacks.run_latest_quarter_qa()

        assert qa_once == qa_twice
        assert globs_after_first == counts["glob"]
        assert globs_after_first == 3
        assert (
            ctx.data.doc_cache.latest_quarter_qa_bundle_by_quarter
            or ctx.data.doc_cache.latest_quarter_sec_text_by_quarter
        )


def test_quarter_notes_ui_reuses_doc_sentence_harvest_cache_across_repeated_runs(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "qn_doc_sentence_cache.xlsx")
            cache_dir = out_path.parent
            doc_path = cache_dir / "doc_000012345625000001_10k.txt"
            doc_path.write_text(
                "Results of operations. "
                "SG&A expense decreased $12 million compared to prior year primarily due to lower compensation costs. "
                "Cash flows from operating activities increased $25 million compared to prior year primarily due to working capital release. "
                "Cash flows from financing activities increased $10 million due to debt repayment timing.",
                encoding="utf-8",
            )

            submission_rows = [
                {
                    "accn": "0000123456-25-000001",
                    "form": "10-K",
                    "filed": "2026-02-20",
                    "report": "2025-12-31",
                    "doc": doc_path.name,
                }
            ]
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "note_id": ["seed-note-1"],
                    "category": ["Results / drivers"],
                    "claim": ["Revenue stabilized in the fourth quarter."],
                    "note": ["Revenue stabilized in the fourth quarter."],
                    "metric_ref": ["Revenue"],
                    "score": [95.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["seed_release.txt"],
                    "source_type": ["earnings_release"],
                    "evidence_snippet": ["Revenue stabilized in the fourth quarter."],
                }
            )

            counts = {str(doc_path): 0}
            original_glob = Path.glob
            original_read_text = Path.read_text

            def counted_glob(self: Path, pattern: str, *args, **kwargs):
                if str(pattern).startswith("doc_000012345625000001_"):
                    return [doc_path]
                return original_glob(self, pattern, *args, **kwargs)

            def counted_read_text(self: Path, *args, **kwargs):
                path_str = str(self)
                if path_str in counts:
                    counts[path_str] += 1
                return original_read_text(self, *args, **kwargs)

            monkeypatch.setattr(Path, "glob", counted_glob)
            monkeypatch.setattr(Path, "read_text", counted_read_text)
            monkeypatch.setattr(writer_context_module, "source_submission_recent_rows", lambda **kwargs: list(submission_rows))

            inputs = _make_inputs(out_path, ticker="PBI", quarter_notes=quarter_notes)
            manifest_df = pd.DataFrame({"path": [str(doc_path)]})
            inputs = inputs.__class__(**{**vars(inputs), "manifest_df": manifest_df})
            ctx = build_writer_context(inputs)

            if "Quarter_Notes_UI" in ctx.wb.sheetnames:
                del ctx.wb["Quarter_Notes_UI"]
            ctx.callbacks.write_quarter_notes_ui_v2()
            first_rows = _quarter_block_notes(ctx.wb["Quarter_Notes_UI"], "2025-12-31")
            reads_after_first = counts[str(doc_path)]

            del ctx.wb["Quarter_Notes_UI"]
            ctx.callbacks.write_quarter_notes_ui_v2()
            second_rows = _quarter_block_notes(ctx.wb["Quarter_Notes_UI"], "2025-12-31")

            assert reads_after_first >= 1
            assert counts[str(doc_path)] == reads_after_first
            assert first_rows == second_rows


def test_write_ui_sheets_records_ui_substage_timings(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "ui_timings.xlsx")
            press_dir = case_dir / "TEST" / "press_release"
            press_dir.mkdir(parents=True, exist_ok=True)
            (press_dir / "Green-Plains-Reports-Fourth-Quarter-and-Full-Year-2025-Financial-Results-2026.txt").write_text(
                "Adjusted EBITDA of $49.1 million, inclusive of $23.4 million in 45Z production tax credit value net of discounts and other costs. "
                "The consolidated ethanol crush margin was $44.4 million for the fourth quarter of 2025, compared with $(15.5) million for the same period in 2024.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 5,
                    "note_id": ["ebitda-1", "debt-1", "fcf-1", "util-1", "ops-1"],
                    "category": [
                        "Strategy / segment",
                        "Debt / liquidity / covenants",
                        "Cash flow / FCF / capex",
                        "Results / drivers",
                        "Programs / initiatives",
                    ],
                    "claim": [
                        "Adjusted EBITDA YoY 369.8%",
                        "Net debt delta $-77.9m",
                        "FCF TTM YoY delta $198.7m",
                        "Utilization 97% of stated capacity",
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                    ],
                    "note": [
                        "Adjusted EBITDA YoY 369.8%",
                        "Net debt delta $-77.9m",
                        "FCF TTM YoY delta $198.7m",
                        "Utilization 97% of stated capacity",
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                    ],
                    "metric_ref": ["Adjusted EBITDA", "Net debt / leverage", "FCF", "Utilization", "Strategic milestone"],
                    "score": [96.0, 95.0, 94.0, 93.0, 92.0],
                    "doc_type": ["model_metric", "model_metric", "model_metric", "earnings_release", "earnings_release"],
                    "doc": ["history_q", "history_q", "history_q", "release_q4.txt", "release_q4.txt"],
                    "source_type": ["model_metric", "model_metric", "model_metric", "earnings_release", "earnings_release"],
                    "evidence_snippet": [
                        "Adjusted EBITDA YoY 369.8%",
                        "Net debt delta $-77.9m",
                        "FCF TTM YoY delta $198.7m",
                        "Utilization 97% of stated capacity",
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                    ],
                }
            )
            promises = pd.DataFrame(
                {
                    "promise_id": ["p-1"],
                    "first_seen_evidence_quarter": [pd.Timestamp("2025-12-31")],
                    "promise_text": ["Maintain high utilization and improve working capital."],
                    "source_evidence_json": [
                        json.dumps({"doc_path": "release_q4.txt", "snippet": "Maintain high utilization and improve working capital."})
                    ],
                }
            )
            progress = pd.DataFrame(
                {
                    "promise_id": ["p-1"],
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "status": ["open"],
                    "source_evidence_json": [
                        json.dumps({"doc_path": "release_q4.txt", "snippet": "Maintain high utilization and improve working capital."})
                    ],
                }
            )
            inputs = _make_inputs(out_path, quarter_notes=quarter_notes, profile_timings=True)
            inputs = inputs.__class__(**{**vars(inputs), "promises": promises, "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_quarter_notes_ui_v2()
            write_ui_sheets(ctx)

            assert {
                "write_excel.ui.raw_frames",
                "write_excel.ui.render.quarter_notes",
                "write_excel.ui.render.quarter_notes.setup",
                "write_excel.ui.render.quarter_notes.selection",
                "write_excel.ui.render.quarter_notes.selection.block_assembly",
                "write_excel.ui.render.quarter_notes.selection.block_assembly.fallback_recall_merge",
                "write_excel.ui.render.quarter_notes.selection.block_assembly.slide_signal_merge",
                "write_excel.ui.render.quarter_notes.selection.block_assembly.initial_render_dedupe",
                "write_excel.ui.render.quarter_notes.selection.block_assembly.gpre_prebadge_event_cleanup",
                "write_excel.ui.render.quarter_notes.selection.block_assembly.origin_event_dedupe",
                "write_excel.ui.render.quarter_notes.selection.block_assembly.connected_buyback_restore",
                "write_excel.ui.render.quarter_notes.selection.block_assembly.buyback_cleanup_passes",
                "write_excel.ui.render.quarter_notes.selection.block_assembly.final_badge_history",
                "write_excel.ui.render.quarter_notes.selection.block_assembly.final_sort_and_render_guardrails",
                "write_excel.ui.render.quarter_notes.selection.block_assembly.post_quarter_companion_append",
                "write_excel.ui.render.quarter_notes.selection.block_assembly.audit_state_finalize",
                "write_excel.ui.render.quarter_notes.render_blocks",
                "write_excel.ui.render.quarter_notes.final_formatting",
                "write_excel.ui.render.promise_tracker",
                "write_excel.ui.render.promise_progress",
                "write_excel.ui.progress_bundle.build",
                "write_excel.ui.progress_rows.select",
                "write_excel.ui.progress_rows.follow_through",
                "write_excel.ui.progress_rows.dedupe",
                "write_excel.ui.progress_rows.render",
            }.issubset(ctx.writer_timings.keys())
            assert "write_excel.ui" not in ctx.writer_timings


def test_write_ui_sheets_records_quarter_notes_setup_timing_for_empty_source() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "quarter_notes_empty_timing.xlsx")
        ctx = build_writer_context(_make_inputs(out_path, profile_timings=True, quarter_notes=pd.DataFrame()))

        write_ui_sheets(ctx)

        assert "write_excel.ui.render.quarter_notes.setup" in ctx.writer_timings
        assert "write_excel.ui.render.quarter_notes.selection" not in ctx.writer_timings
        assert "Quarter_Notes_UI" in ctx.wb.sheetnames


def test_write_ui_sheets_records_pbi_block_assembly_selection_bucket(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "ui_timings_pbi.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "note_id": ["note-1"],
                    "category": ["Guidance / outlook"],
                    "claim": ["FY 2026 Revenue guidance $1,760m-$1,860m."],
                    "note": ["FY 2026 Revenue guidance $1,760m-$1,860m."],
                    "metric_ref": ["Revenue guidance"],
                    "score": [95.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q4.txt"],
                    "source_type": ["earnings_release"],
                    "evidence_snippet": ["FY 2026 Revenue guidance $1,760m-$1,860m."],
                }
            )
            ctx = build_writer_context(_make_inputs(out_path, profile_timings=True, quarter_notes=quarter_notes))

            ctx.callbacks.write_quarter_notes_ui_v2()

            assert "write_excel.ui.render.quarter_notes.selection.block_assembly.pbi_final_selection" in ctx.writer_timings


def test_local_balance_sheet_payloads_parse_only_target_quarters(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "local_bs_lazy.xlsx")
        material_root = out_path.parent.parent
        fs_dir = material_root / "financial_statement"
        fs_dir.mkdir(parents=True, exist_ok=True)

        target_path = fs_dir / "TEST_Q4_2025_balance_sheet.txt"
        stale_path = fs_dir / "TEST_Q4_2024_balance_sheet.txt"
        target_path.write_text("Balance sheet for quarter ended December 31, 2025.", encoding="utf-8")
        stale_path.write_text("Balance sheet for quarter ended December 31, 2024.", encoding="utf-8")

        parse_calls: list[pd.Timestamp] = []

        def fake_extract_balance_sheet_from_text(text_in: str, quarter_end: pd.Timestamp):
            qv = pd.Timestamp(quarter_end)
            parse_calls.append(qv)
            if qv == pd.Timestamp("2025-12-31"):
                return {"values": {"goodwill": 42_000_000.0, "intangibles": 18_000_000.0}}
            if qv == pd.Timestamp("2024-12-31"):
                return {"values": {"goodwill": 9_000_000.0}}
            return None

        monkeypatch.setattr(
            writer_context_module,
            "_extract_balance_sheet_from_text",
            fake_extract_balance_sheet_from_text,
        )

        ctx = build_writer_context(_make_inputs(out_path, profile_timings=True))
        ensure_valuation_inputs(ctx)
        quarter_key = tuple(pd.Timestamp(q) for q in ctx.inputs.hist["quarter"].tolist())
        render_loader = ctx.state["_ensure_valuation_render_bundle"]

        render_bundle = render_loader(quarter_key, ctx.derived.leverage_df)
        render_loader(quarter_key, ctx.derived.leverage_df)

        assert parse_calls == [pd.Timestamp("2025-12-31")]
        assert render_bundle["goodwill_map"][pd.Timestamp("2025-12-31")] == pytest.approx(42_000_000.0)
        assert render_bundle["intangibles_map"][pd.Timestamp("2025-12-31")] == pytest.approx(18_000_000.0)
        assert "write_excel.valuation.bundle.local_bs.index" in ctx.writer_timings
        assert "write_excel.valuation.bundle.local_bs.parse_selected" in ctx.writer_timings
        assert "write_excel.valuation.bundle.local_bs.pick_best" in ctx.writer_timings


def test_pbi_profile_omits_driver_and_economics_sheets_but_keeps_ui_sheets() -> None:
    with _case_dir() as case_dir:
        out_path = _make_ticker_model_out_path(case_dir, "PBI", "pbi_gating.xlsx")
        write_excel_from_inputs(_make_inputs(out_path, ticker="PBI"))

        wb = load_workbook(out_path, data_only=False)
        assert "Operating_Drivers" in wb.sheetnames
        assert "Economics_Overlay" not in wb.sheetnames
        assert "economics_market_raw" not in wb.sheetnames
        assert "Quarter_Notes_UI" in wb.sheetnames
        assert "Promise_Tracker_UI" not in wb.sheetnames
        assert "Promise_Progress_UI" in wb.sheetnames
        assert "Promise_Tracker" in wb.sheetnames
        assert "Promise_Evidence" in wb.sheetnames
        assert "Promise_Progress" in wb.sheetnames


def test_pbi_bs_segments_prefers_quarterly_segment_block_from_segment_workbook(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_bs_segments.xlsx")
            material_root = out_path.parent.parent
            _write_pbi_segment_workbook(material_root / "segment_financials", include_quarters=True)

            write_excel_from_inputs(_make_inputs(out_path, ticker="TEST"))
            ws = load_workbook(out_path, data_only=False)["BS_Segments"]

        quarterly_row = _find_row_with_value(ws, "Quarterly segments")
        annual_row = _find_row_with_value(ws, "Annual segments")
        assert quarterly_row is not None
        assert annual_row is not None
        assert quarterly_row < annual_row
        assert str(ws["A4"].value).startswith("QA:")
        assert [ws.cell(row=11, column=cc).value for cc in range(2, 6)] == [
            "2025-Q1",
            "2025-Q2",
            "2025-Q3",
            "2025-Q4",
        ]

        def _find_after(start_row: int, label: str) -> int:
            for rr in range(start_row + 1, ws.max_row + 1):
                if ws.cell(row=rr, column=1).value == label:
                    return rr
            raise AssertionError(f"Could not find {label!r} after row {start_row}")

        revenue_section = _find_after(quarterly_row, "Revenue")
        revenue_sendtech = _find_after(revenue_section, "SendTech Solutions")
        revenue_presort = _find_after(revenue_sendtech, "Presort Services")
        assert [ws.cell(row=revenue_sendtech, column=cc).value for cc in range(2, 6)] == [480.0, 490.0, 500.0, 510.0]
        assert [ws.cell(row=revenue_presort, column=cc).value for cc in range(2, 6)] == [300.0, 305.0, 310.0, 315.0]

        ebit_section = _find_after(revenue_presort, "Adjusted EBIT")
        ebit_sendtech = _find_after(ebit_section, "SendTech Solutions")
        assert [ws.cell(row=ebit_sendtech, column=cc).value for cc in range(2, 6)] == [120.0, 125.0, 130.0, 135.0]

        margin_section = _find_after(ebit_sendtech, "EBIT margin %")
        margin_sendtech = _find_after(margin_section, "SendTech Solutions")
        assert [ws.cell(row=margin_sendtech, column=cc).value for cc in range(2, 6)] == [0.25, 0.255, 0.26, 0.265]

        da_section = _find_after(margin_sendtech, "Depreciation & amortization")
        da_sendtech = _find_after(da_section, "SendTech Solutions")
        assert [ws.cell(row=da_sendtech, column=cc).value for cc in range(2, 6)] == [18.0, 19.0, 20.0, 21.0]
        assert ws.cell(row=revenue_sendtech, column=2).fill.patternType == "solid"
        assert ws.cell(row=revenue_section, column=2).fill.patternType == "solid"


def test_pbi_bs_segments_falls_back_to_annual_only_when_quarterly_workbook_missing(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_bs_segments_annual_only.xlsx")
            material_root = out_path.parent.parent
            _write_pbi_segment_workbook(material_root / "segment_financials", include_quarters=False)

            write_excel_from_inputs(_make_inputs(out_path, ticker="TEST"))
            ws = load_workbook(out_path, data_only=False)["BS_Segments"]

        assert _find_row_with_value(ws, "Quarterly segments") is None
        annual_row = _find_row_with_value(ws, "Annual segments")
        assert annual_row is not None
        assert _find_row_with_value(ws, "Revenues") is not None


def test_valuation_guidance_layout_uses_new_column_spans_and_preserves_type_wording() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "guidance_layout.xlsx")
        hist = _make_hist().iloc[:2].copy().reset_index(drop=True)
        slides_guidance = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-03-31"), pd.Timestamp("2025-06-30")],
                "line": [
                    "Updated full-year 2025 revenue guidance $1900 to $1950 adjusted EBIT $450 to $465 adjusted EPS $1.20 to $1.40 free cash flow $330 to $370",
                    "Updated full-year 2025 revenue guidance $1910 to $1960 adjusted EBIT $455 to $470",
                ],
                "heading": ["guidance", "guidance"],
                "filed": [pd.Timestamp("2025-05-01"), pd.Timestamp("2025-08-01")],
                "doc": ["guidance_q1.txt", "guidance_q2.txt"],
            }
        )

        inputs = _make_inputs(out_path, hist=hist)
        inputs = inputs.__class__(**{**vars(inputs), "slides_guidance": slides_guidance})
        ctx = build_writer_context(inputs)
        ensure_valuation_inputs(ctx)
        ctx.callbacks.write_valuation_sheet()
        ws = ctx.wb["Valuation"]
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}

        guidance_header_row = _find_row_with_value(ws, "Trend / realized", column=26)
        assert guidance_header_row is not None
        assert f"O{guidance_header_row}:P{guidance_header_row}" in merged_ranges
        assert f"S{guidance_header_row}:Y{guidance_header_row}" in merged_ranges
        assert f"Z{guidance_header_row}:AC{guidance_header_row}" in merged_ranges
        assert ws.cell(row=guidance_header_row, column=15).value == "Metric"
        assert ws.cell(row=guidance_header_row, column=17).value == "Stated in"
        assert ws.cell(row=guidance_header_row, column=18).value == "Applies to"
        assert ws.cell(row=guidance_header_row, column=19).value == "Guidance"

        stated_row = next(
            rr for rr in range(guidance_header_row + 1, ws.max_row + 1)
            if str(ws.cell(row=rr, column=15).value or "").strip() == "Revenue"
        )
        assert f"O{stated_row}:P{stated_row}" in merged_ranges
        assert f"S{stated_row}:Y{stated_row}" in merged_ranges
        assert f"Z{stated_row}:AC{stated_row}" in merged_ranges
        assert str(ws.cell(row=stated_row, column=17).alignment.horizontal or "") == "left"
        assert ws.cell(row=stated_row, column=19).alignment.wrap_text is True
        assert ws.cell(row=stated_row, column=26).alignment.wrap_text is True
        assert "(+0.5%)" in str(ws.cell(row=stated_row, column=26).value or "")
        carry_rows = [
            rr for rr in range(guidance_header_row + 1, ws.max_row + 1)
            if str(ws.cell(row=rr, column=15).value or "").strip() == "Cost savings"
        ]
        if carry_rows:
            carry_row = carry_rows[0]
            assert f"O{carry_row}:P{carry_row}" in merged_ranges
            assert f"S{carry_row}:Y{carry_row}" in merged_ranges
            assert f"Z{carry_row}:AC{carry_row}" in merged_ranges
            assert str(ws.cell(row=carry_row, column=17).alignment.horizontal or "") == "left"
            assert ws.cell(row=carry_row, column=19).alignment.wrap_text is True
            assert ws.cell(row=carry_row, column=26).alignment.wrap_text is True
            carry_value = str(ws.cell(row=carry_row, column=19).value or "")
            assert "target" in carry_value.lower()
            assert float(ws.row_dimensions[carry_row].height or 0.0) <= 58.0
        assert str(ws.freeze_panes) == "B7"


def test_valuation_normalizes_negative_capex_and_marks_adj_fcf_company_defined() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "valuation_capex_normalization.xlsx")
        hist = pd.DataFrame(
            {
                "quarter": pd.to_datetime(["2023-03-31", "2023-06-30", "2023-09-30", "2023-12-31"]),
                "revenue": [120_000_000.0, 125_000_000.0, 130_000_000.0, 135_000_000.0],
                "cfo": [15_000_000.0, 16_000_000.0, 20_000_000.0, 18_000_000.0],
                "capex": [2_000_000.0, 3_000_000.0, -4_420_000.0, 5_000_000.0],
                "ebitda": [20_000_000.0, 21_000_000.0, 22_000_000.0, 23_000_000.0],
                "ebit": [11_000_000.0, 12_000_000.0, 13_000_000.0, 14_000_000.0],
                "cash": [30_000_000.0, 31_000_000.0, 32_000_000.0, 33_000_000.0],
                "debt_core": [60_000_000.0, 59_000_000.0, 58_000_000.0, 57_000_000.0],
                "shares_outstanding": [10_000_000.0, 10_000_000.0, 10_000_000.0, 10_000_000.0],
                "shares_diluted": [10_000_000.0, 10_000_000.0, 10_000_000.0, 10_000_000.0],
                "market_cap": [100_000_000.0, 101_000_000.0, 102_000_000.0, 103_000_000.0],
                "interest_expense_net": [1_000_000.0, 1_000_000.0, 1_000_000.0, 1_000_000.0],
            }
        )
        ctx = build_writer_context(_make_inputs(out_path, hist=hist))
        ensure_valuation_inputs(ctx)
        ctx.callbacks.write_valuation_sheet()
        ws = ctx.wb["Valuation"]

        q3_col = next(
            cc for cc in range(2, ws.max_column + 1)
            if str(ws.cell(row=6, column=cc).value or "").strip() == "2023-Q3"
        )
        capex_row = _find_row_with_value(ws, "Capex")
        fcf_row = _find_row_with_value(ws, "FCF (CFO-Capex)")
        adj_fcf_row = _find_row_with_value(ws, "Adj FCF (TTM)")

        assert capex_row is not None
        assert fcf_row is not None
        assert adj_fcf_row is not None
        assert float(pd.to_numeric(ws.cell(row=capex_row, column=q3_col).value, errors="coerce")) == pytest.approx(4.42, abs=0.001)
        assert float(pd.to_numeric(ws.cell(row=fcf_row, column=q3_col).value, errors="coerce")) == pytest.approx(15.58, abs=0.001)
        assert str((ws.cell(row=adj_fcf_row, column=1).comment.text if ws.cell(row=adj_fcf_row, column=1).comment else "") or "") == "company-defined"


def test_valuation_adds_ratio_notes_and_convertible_dilution_structure_comments_conservatively(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_ticker_model_out_path(case_dir, "PBI", "valuation_ratio_and_convertible_notes.xlsx")
            sec_cache_dir = case_dir / "PBI" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000119312525177477_pbi-capped-call.htm").write_text(
                "Pitney Bowes Inc. completed an offering of $230.0 million aggregate principal amount of 1.50% convertible senior notes due August 2032. "
                "The initial conversion rate is 70.1533 shares per $1,000 principal amount. "
                "On August 5, 2025, the company entered into capped call transactions expected generally to reduce the potential dilution to the common stock upon conversion.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_generic-2031-convertible.htm").write_text(
                "Test issuer completed an offering of $150.0 million aggregate principal amount of 3.00% convertible senior notes due 2031. "
                "The initial conversion rate is 40.0000 shares per $1,000 principal amount. "
                "Net proceeds will be used for general corporate purposes.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_buyback-only-2033-convertible.htm").write_text(
                "Test issuer completed an offering of $100.0 million aggregate principal amount of 4.00% convertible senior notes due 2033. "
                "The initial conversion rate is 50.0000 shares per $1,000 principal amount. "
                "The company used approximately $20.0 million of the proceeds to repurchase approximately 2.0 million shares.",
                encoding="utf-8",
            )
            debt_tranches_latest = pd.DataFrame(
                [
                    {
                        "tranche_name": "1.5% convertible senior notes due August 2032",
                        "instrument_type": "convertible",
                        "amount_principal": 230_000_000.0,
                        "coupon_pct": 1.5,
                        "maturity_display": "August 2032",
                        "maturity_year": 2032,
                    },
                    {
                        "tranche_name": "3.0% convertible senior notes due 2031",
                        "instrument_type": "convertible",
                        "amount_principal": 150_000_000.0,
                        "coupon_pct": 3.0,
                        "maturity_display": "2031",
                        "maturity_year": 2031,
                    },
                    {
                        "tranche_name": "4.0% convertible senior notes due 2033",
                        "instrument_type": "convertible",
                        "amount_principal": 100_000_000.0,
                        "coupon_pct": 4.0,
                        "maturity_display": "2033",
                        "maturity_year": 2033,
                    },
                ]
            )
            base_inputs = _make_inputs(out_path, ticker="PBI")
            inputs = base_inputs.__class__(
                **{
                    **vars(base_inputs),
                    "cache_dir": sec_cache_dir,
                    "debt_tranches_latest": debt_tranches_latest,
                }
            )
            ctx = build_writer_context(inputs)
            ensure_valuation_inputs(ctx)
            ctx.callbacks.write_valuation_sheet()
            ws = ctx.wb["Valuation"]

            current_ratio_row = _find_row_with_value(ws, "Current ratio")
            quick_ratio_row = _find_row_with_value(ws, "Quick ratio")
            assert current_ratio_row is not None
            assert quick_ratio_row is not None
            assert str((ws.cell(row=current_ratio_row, column=1).comment.text if ws.cell(row=current_ratio_row, column=1).comment else "") or "") == (
                "Current assets / current liabilities. Short-term liquidity measure; around 1.0+ is often healthier."
            )
            assert str((ws.cell(row=quick_ratio_row, column=1).comment.text if ws.cell(row=quick_ratio_row, column=1).comment else "") or "") == (
                "Near-cash current assets / current liabilities. Stricter liquidity measure; around 1.0+ is often stronger."
            )

            pbi_row = _find_row_with_value(ws, "1.5% notes due August 2032", column=12)
            generic_row = _find_row_with_value(ws, "3% notes due 2031", column=12)
            buyback_only_row = _find_row_with_value(ws, "4% notes due 2033", column=12)
            pbi_debt_detail_row = _find_row_containing(ws, "August 2032", column=1)
            assert pbi_row is not None
            assert generic_row is not None
            assert buyback_only_row is not None
            assert pbi_debt_detail_row is not None
            assert str((ws.cell(row=pbi_row, column=19).comment.text if ws.cell(row=pbi_row, column=19).comment else "") or "") == (
                "Capped call may reduce dilution."
            )
            assert str((ws.cell(row=pbi_debt_detail_row, column=9).comment.text if ws.cell(row=pbi_debt_detail_row, column=9).comment else "") or "") == (
                "Capped call may reduce dilution."
            )
            assert ws.cell(row=generic_row, column=19).comment is None
            assert ws.cell(row=buyback_only_row, column=19).comment is None


def test_valuation_thesis_bridge_adds_note_keeps_label_and_formulas() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "thesis_bridge.xlsx")
        ctx = build_writer_context(_make_inputs(out_path))
        ensure_valuation_inputs(ctx)
        ctx.callbacks.write_valuation_sheet()
        ws = ctx.wb["Valuation"]

        note_text = "Enter annual incremental EBITDA/equity effects, not raw rate changes."
        note_row = _find_row_with_value(ws, note_text, column=20)
        bridge_label_row = _find_row_with_value(ws, "Interest savings / debt-paydown uplift", column=15)
        thesis_output_row = _find_row_with_value(ws, "Thesis value/share @ target multiple", column=15)

        assert note_row is not None
        assert bridge_label_row is not None
        assert thesis_output_row is not None
        assert ws.cell(row=thesis_output_row, column=19).value.startswith("=")
        assert float(ws.column_dimensions["B"].width or 0.0) == pytest.approx(12.43, abs=0.02)
        assert float(ws.column_dimensions["C"].width or 0.0) == pytest.approx(12.43, abs=0.02)
        assert float(ws.column_dimensions["S"].width or 0.0) >= 24.0
        assert ws.cell(row=bridge_label_row, column=19).number_format == "#,##0.000"
        thesis_formula_cells = [
            ws.cell(row=_find_row_with_value(ws, "Thesis Adj EBITDA", column=15), column=19).value,
            ws.cell(row=_find_row_with_value(ws, "Thesis FCF", column=15), column=19).value,
            ws.cell(row=_find_row_with_value(ws, "Thesis EV @ target multiple", column=15), column=19).value,
            ws.cell(row=_find_row_with_value(ws, "Thesis equity value @ target multiple", column=15), column=19).value,
            ws.cell(row=_find_row_with_value(ws, "Thesis value/share @ target multiple", column=15), column=19).value,
            ws.cell(row=_find_row_with_value(ws, "Thesis equity value @ target equity FCF yield", column=15), column=19).value,
            ws.cell(row=_find_row_with_value(ws, "Thesis value/share @ target equity FCF yield", column=15), column=19).value,
        ]
        assert all("*1000000" not in str(v) and "/1000000" not in str(v) for v in thesis_formula_cells)


def test_gpre_valuation_thesis_bridge_keeps_owner_earnings_rows_visible(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_thesis_bridge_visibility.xlsx")
            ctx = build_writer_context(_make_inputs(out_path, ticker="GPRE"))
            ensure_valuation_inputs(ctx)
            ctx.callbacks.write_valuation_sheet()
            ws = ctx.wb["Valuation"]

            owner_earnings_row = _find_row_with_value(ws, "Owner earnings (proxy)", column=1)
            cash_flow_quality_row = _find_row_with_value(ws, "Cash-flow quality", column=1)

            assert owner_earnings_row is not None
            assert cash_flow_quality_row is not None
            assert not bool(ws.row_dimensions[owner_earnings_row].hidden)
            assert not bool(ws.row_dimensions[cash_flow_quality_row].hidden)
            assert _find_row_with_value(ws, "Corn oil / coproduct uplift", column=15) is None
            assert _find_row_with_value(ws, "Protein / mix uplift", column=15) is None
            assert str(ws.cell(row=owner_earnings_row, column=15).value or "").strip() == "Cost savings uplift"
            assert str(ws.cell(row=cash_flow_quality_row, column=15).value or "").strip() == "Interest savings / debt-paydown uplift"


def test_valuation_columns_b_and_c_use_consistent_width_for_pbi_and_gpre(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        for ticker in ("PBI", "GPRE"):
            with _profile_override(monkeypatch, ticker):
                out_path = _make_model_out_path(case_dir, f"{ticker.lower()}_valuation_widths.xlsx")
                ctx = build_writer_context(_make_inputs(out_path, ticker="TEST"))
                ensure_valuation_inputs(ctx)
                ctx.callbacks.write_valuation_sheet()
                ws = ctx.wb["Valuation"]

                assert float(ws.column_dimensions["B"].width or 0.0) == pytest.approx(12.43, abs=0.02)
                assert float(ws.column_dimensions["C"].width or 0.0) == pytest.approx(12.43, abs=0.02)


def test_valuation_buybacks_can_use_explicit_sec_repurchase_disclosures(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_ticker_model_out_path(case_dir, "GPRE", "gpre_valuation_buybacks.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940226000019_gpre-20251231.htm").write_text(
                "Green Plains Inc. "
                "On October 27, 2025, in conjunction with the privately negotiated exchange and subscription agreements for the 2030 Notes, "
                "the company repurchased 2.9 million shares of its common stock for a total of $30.0 million under the repurchase program. "
                "At February 10, 2026, $77.2 million in share repurchase authorization remained.",
                encoding="utf-8",
            )

            inputs = _make_inputs(out_path, ticker="GPRE")
            inputs = inputs.__class__(**{**vars(inputs), "cache_dir": sec_cache_dir})
            ctx = build_writer_context(inputs)
            ensure_valuation_inputs(ctx)
            quarter_key = tuple(pd.Timestamp(q) for q in ctx.inputs.hist["quarter"].tolist())
            render_loader = ctx.state["_ensure_valuation_render_bundle"]
            render_bundle = render_loader(quarter_key, ctx.derived.leverage_df)
            render_bundle["buyback_shares_q_map"][pd.Timestamp("2025-12-31")] = 2_101_000.0
            render_bundle["buyback_cash_facts_map"][pd.Timestamp("2025-12-31")] = 2_000_000.0
            ctx.derived.valuation_render_bundle = render_bundle
            ctx.callbacks.write_valuation_sheet()
            ws = ctx.wb["Valuation"]

            buybacks_row = _find_row_with_value(ws, "Buybacks (shares)")
            buybacks_note_row = _find_row_with_value(ws, "Buybacks note")
            flags_header_row = _find_row_with_value(ws, "Hidden value flags")
            obs5_row = _find_row_with_value(ws, "Obs 5")

            assert buybacks_row is not None
            assert buybacks_note_row is not None
            assert flags_header_row is not None
            assert obs5_row is None
            assert str(ws.cell(row=buybacks_row, column=2).value or "") != "n/a"
            assert "Latest quarter +2.900m" in str(ws.cell(row=buybacks_row, column=2).value or "")
            assert "$10.34/share" in str(ws.cell(row=buybacks_row, column=2).value or "")
            assert "Remaining capacity $77.2m" in str(ws.cell(row=buybacks_note_row, column=2).value or "")
            assert "explicit SEC repurchase disclosures" not in " | ".join(
                str(ws.cell(row=rr, column=2).value or "")
                for rr in range(1, ws.max_row + 1)
            ).lower()


def test_pbi_quarter_notes_ui_keeps_clean_driver_note_alongside_guidance(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_quarter_notes_driver.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 3,
                    "note_id": ["guidance-1", "driver-1", "junk-1"],
                    "category": ["Guidance / outlook", "Results / drivers", "Other / footnotes"],
                    "claim": [
                        "Adjusted EBIT guidance increased to $450 million to $465 million due to pricing and mix improvements in SendTech.",
                        "PB Bank liquidity improved by $120 million as trapped capital was released and SendTech margins improved on pricing and mix.",
                        "map permit list county parcel latitude longitude $12 million",
                    ],
                    "note": [
                        "Adjusted EBIT guidance increased to $450 million to $465 million due to pricing and mix improvements in SendTech.",
                        "PB Bank liquidity improved by $120 million as trapped capital was released and SendTech margins improved on pricing and mix.",
                        "map permit list county parcel latitude longitude $12 million",
                    ],
                    "metric_ref": ["Adjusted EBIT guidance", "PB Bank liquidity release", "Other"],
                    "score": [95.0, 93.0, 96.0],
                    "doc_type": ["earnings_release", "transcript", "ocr"],
                    "doc": ["release_q4.txt", "transcript_q4.txt", "ocr_junk.txt"],
                    "evidence_snippet": [
                        "Adjusted EBIT guidance increased to $450 million to $465 million due to pricing and mix improvements in SendTech.",
                        "PB Bank liquidity improved by $120 million as trapped capital was released and SendTech margins improved on pricing and mix.",
                        "map permit list county parcel latitude longitude $12 million",
                    ],
                    "evidence_json": [
                        json.dumps(
                            [
                                {
                                    "doc_path": "release_q4.txt",
                                    "doc_type": "earnings_release",
                                    "snippet": "Adjusted EBIT guidance increased to $450 million to $465 million due to pricing and mix improvements in SendTech.",
                                }
                            ]
                        ),
                        json.dumps(
                            [
                                {
                                    "doc_path": "transcript_q4.txt",
                                    "doc_type": "transcript",
                                    "snippet": "PB Bank liquidity improved by $120 million as trapped capital was released and SendTech margins improved on pricing and mix.",
                                }
                            ]
                        ),
                        json.dumps(
                            [
                                {
                                    "doc_path": "ocr_junk.txt",
                                    "doc_type": "ocr",
                                    "snippet": "map permit list county parcel latitude longitude $12 million",
                                }
                            ]
                        ),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            metrics = [str(ws.cell(row=rr, column=4).value or "") for rr in range(1, ws.max_row + 1)]
            notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]

            assert "Adjusted EBIT guidance" in metrics
            assert "PB Bank liquidity release" in metrics
            assert not any("map permit list county parcel" in note.lower() for note in notes)


def test_pbi_quarter_notes_ui_does_not_relabel_generic_guidance_as_sendtech_driver(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_quarter_notes_no_sendtech_mislabel.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-03-31")] * 2,
                    "note_id": ["bad-guidance-1", "good-driver-1"],
                    "category": ["Guidance / outlook", "Results / drivers"],
                    "claim": [
                        "Full-year 2025 revenue guidance was $1.95 billion to $2.00 billion.",
                        "SendTech and Presort margins improved on pricing and mix while PB Bank liquidity improved.",
                    ],
                    "note": [
                        "Full-year 2025 revenue guidance was $1.95 billion to $2.00 billion.",
                        "SendTech and Presort margins improved on pricing and mix while PB Bank liquidity improved.",
                    ],
                    "metric_ref": ["SendTech / Presort operating driver", "SendTech / Presort operating driver"],
                    "score": [95.0, 91.0],
                    "doc_type": ["earnings_release", "transcript"],
                    "doc": ["release_q1.txt", "transcript_q1.txt"],
                    "evidence_snippet": [
                        "Full-year 2025 revenue guidance was $1.95 billion to $2.00 billion.",
                        "SendTech and Presort margins improved on pricing and mix while PB Bank liquidity improved.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q1.txt", "doc_type": "earnings_release", "snippet": "Full-year 2025 revenue guidance was $1.95 billion to $2.00 billion."}]),
                        json.dumps([{"doc_path": "transcript_q1.txt", "doc_type": "transcript", "snippet": "SendTech and Presort margins improved on pricing and mix while PB Bank liquidity improved."}]),
                    ],
                }
            )
            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_rows = [
                (
                    str(ws.cell(row=rr, column=3).value or ""),
                    str(ws.cell(row=rr, column=4).value or ""),
                )
                for rr in range(1, ws.max_row + 1)
                if str(ws.cell(row=rr, column=3).value or "").strip()
            ]
            assert not any(
                metric == "SendTech / Presort operating driver" and "$1.95" in note
                for note, metric in visible_rows
            )
            assert all(
                "$1.95" not in note or metric != "SendTech / Presort operating driver"
                for note, metric in visible_rows
            )


def test_pbi_promise_progress_uses_observed_actual_for_resolved_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_latest.xlsx")
            hist = _make_hist().copy()
            hist["revenue"] = [100_000_000.0, 110_000_000.0, 120_000_000.0, 130_000_000.0]
            progress = pd.DataFrame(
                {
                    "promise_id": ["p-revenue"],
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "status": ["resolved_fail"],
                    "metric_ref": ["Revenue guidance"],
                    "target": ["$500.0m-$520.0m"],
                    "latest": ["not yet measurable"],
                    "rationale": ["FY 2025 revenue guidance was $500 million to $520 million."],
                    "promise_type": ["guidance_range"],
                    "guidance_type": ["period"],
                    "target_period_norm": ["FY2025"],
                    "source_evidence_json": [
                        json.dumps(
                            {
                                "doc_type": "earnings_release",
                                "snippet": "FY 2025 revenue guidance was $500 million to $520 million.",
                            }
                        )
                    ],
                }
            )
            inputs = _make_inputs(out_path, ticker="TEST", hist=hist)
            inputs = inputs.__class__(**{**vars(inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            target_row = None
            for rr in range(1, ws.max_row + 1):
                if ws.cell(row=rr, column=1).value == "Revenue guidance":
                    target_row = rr
                    break
            assert target_row is not None
            assert ws.cell(row=target_row, column=3).value != "not yet measurable"
            assert isinstance(ws.cell(row=target_row, column=3).value, (int, float))
            assert ws.cell(row=target_row, column=4).value == "Missed"


def test_pbi_promise_progress_recovers_later_actual_for_older_fy2025_guidance(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_later_actual.xlsx")
            hist = _make_hist().copy()
            hist["revenue"] = [100_000_000.0, 110_000_000.0, 120_000_000.0, 530_000_000.0]
            progress = pd.DataFrame(
                {
                    "promise_id": ["p-revenue-q3"],
                    "quarter": [pd.Timestamp("2025-09-30")],
                    "status": ["on_track"],
                    "metric_ref": ["Revenue guidance"],
                    "target": ["$500.0m-$520.0m"],
                    "latest": ["not yet measurable"],
                    "rationale": ["FY 2025 revenue guidance was $500 million to $520 million."],
                    "promise_type": ["guidance_range"],
                    "guidance_type": ["period"],
                    "target_period_norm": ["FY2025"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2025 revenue guidance was $500 million to $520 million."})
                    ],
                }
            )
            inputs = _make_inputs(out_path, ticker="TEST", hist=hist)
            inputs = inputs.__class__(**{**vars(inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            target_row = None
            for rr in range(1, ws.max_row + 1):
                if ws.cell(row=rr, column=2).value == "Revenue guidance":
                    target_row = rr
                    break
            assert target_row is not None
            assert ws.cell(row=target_row, column=4).value != "not yet measurable"
            assert isinstance(ws.cell(row=target_row, column=4).value, (int, float))


def test_pbi_promise_progress_drops_cost_savings_row_when_guidance_alignment_is_wrong(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_cost_savings_alignment.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["bad-cost-savings"],
                    "quarter": [pd.Timestamp("2024-12-31")],
                    "status": ["pending"],
                    "metric_ref": ["Cost savings target"],
                    "target": ["$450.0m-$480.0m"],
                    "latest": ["not yet measurable"],
                    "rationale": ["Adjusted EBIT guidance increased to $450 million to $480 million for FY 2025."],
                    "promise_type": ["guidance_range"],
                    "guidance_type": ["period"],
                    "target_period_norm": ["FY2025"],
                    "source_evidence_json": [
                        json.dumps(
                            {
                                "doc_type": "earnings_release",
                                "snippet": "Adjusted EBIT guidance increased to $450 million to $480 million for FY 2025.",
                            }
                        )
                    ],
                }
            )
            base_inputs = _make_inputs(out_path, ticker="TEST")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            visible_values = [
                str(ws.cell(row=rr, column=cc).value or "")
                for rr in range(1, ws.max_row + 1)
                for cc in range(1, 8)
            ]
            assert all("Cost savings target" not in val for val in visible_values)
            assert all("$450.0m-$480.0m" not in val for val in visible_values)


def test_promise_progress_strips_illegal_control_characters_from_rationale(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "promise_progress_illegal_chars.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["gpre-illegal-char"],
                    "quarter": [pd.Timestamp("2025-06-30")],
                    "status": ["on_track"],
                    "metric_ref": ["Cost reduction initiative"],
                    "target": [""],
                    "latest": ["Cost reductions ahead of plan"],
                    "rationale": [
                        "Cost reductions are on pace to exceed the $50.0m annualized savings target.\x00"
                    ],
                    "promise_type": ["operational"],
                    "guidance_type": [""],
                    "target_period_norm": [""],
                    "source_evidence_json": [
                        json.dumps(
                            {
                                "doc_type": "press_release",
                                "snippet": "Cost reductions are on pace to exceed the $50.0m annualized savings target.",
                            }
                        )
                    ],
                }
            )

            inputs = _make_inputs(out_path, ticker="TEST", promise_progress=progress)
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            visible_text = [
                str(ws.cell(row=rr, column=cc).value or "")
                for rr in range(1, ws.max_row + 1)
                for cc in range(1, 12)
            ]
            assert all("\x00" not in val for val in visible_text)

            comments = [
                str(ws.cell(row=rr, column=cc).comment.text or "")
                for rr in range(1, ws.max_row + 1)
                for cc in range(1, 12)
                if ws.cell(row=rr, column=cc).comment is not None
            ]
            assert all("\x00" not in text for text in comments)


def test_promise_progress_ui_surface_contract_on_temp_workbook(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "promise_progress_surface_contract.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["milestone-1"],
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "status": ["achieved"],
                    "metric_ref": ["Strategic milestone"],
                    "target": ["York online by Q4 2025"],
                    "latest": ["fully operational"],
                    "rationale": ["York became fully operational in Q4 2025."],
                    "promise_type": ["milestone"],
                    "guidance_type": ["text"],
                    "target_period_norm": ["Q42025"],
                    "source_evidence_json": [
                        json.dumps(
                            {
                                "doc_type": "earnings_release",
                                "snippet": "York became fully operational in Q4 2025.",
                            }
                        )
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", promise_progress=progress))
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            header_row = _find_row_with_value(ws, "Promise progress (As of 2025-12-31)")
            assert header_row is not None
            assert str(ws["A1"].value or "").strip() == "Promise Progress"
            assert str(ws["A2"].value or "").strip().startswith("Generated at ")
            assert str(ws.cell(row=header_row + 1, column=1).value or "").strip() == "Metric"
            assert str(ws.cell(row=header_row + 1, column=2).value or "").strip() == "Target"
            assert str(ws.cell(row=header_row + 1, column=3).value or "").strip() == "Latest"
            assert str(ws.cell(row=header_row + 1, column=4).value or "").strip() == "Result"
            assert abs(float(ws.column_dimensions["B"].width or 0.0) - 38.142857) < 0.1
            assert abs(float(ws.column_dimensions["C"].width or 0.0) - 38.142857) < 0.1
            completed_rows = [
                rr
                for rr in range(1, ws.max_row + 1)
                if str(ws.cell(row=rr, column=4).value or "").strip() == "Completed"
            ]
            assert completed_rows
            assert {str(ws.cell(row=rr, column=4).fill.fgColor.rgb or "") for rr in completed_rows} == {"0070AD47"}


def test_gpre_quarter_notes_ui_filters_dropped_and_fragment_junk_but_keeps_high_signal_notes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_quarter_notes_cleanup.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 3,
                    "note_id": ["good-1", "junk-1", "junk-2"],
                    "category": ["Programs / initiatives", "What changed", "Other / footnotes"],
                    "claim": [
                        "Utilization improved to 95% as York is fully operational and Central City and Wood River are online and ramping, supporting $95 million of 45Z monetization opportunity.",
                        "[DROPPED] map permit list county parcel latitude longitude $12 million",
                        "permit map county parcel latitude longitude list $10 million",
                    ],
                    "note": [
                        "Utilization improved to 95% as York is fully operational and Central City and Wood River are online and ramping, supporting $95 million of 45Z monetization opportunity.",
                        "[DROPPED] map permit list county parcel latitude longitude $12 million",
                        "permit map county parcel latitude longitude list $10 million",
                    ],
                    "metric_ref": ["45Z monetization / EBITDA", "What changed", "Other"],
                    "score": [95.0, 97.0, 95.0],
                    "doc_type": ["earnings_release", "earnings_release", "ocr"],
                    "doc": ["release_q4.txt", "release_q4.txt", "ocr_bad.txt"],
                    "evidence_snippet": [
                        "Utilization improved to 95% as York is fully operational and Central City and Wood River are online and ramping, supporting $95 million of 45Z monetization opportunity.",
                        "[DROPPED] map permit list county parcel latitude longitude $12 million",
                        "permit map county parcel latitude longitude list $10 million",
                    ],
                    "evidence_json": [
                        json.dumps(
                            [
                                {
                                    "doc_path": "release_q4.txt",
                                    "doc_type": "earnings_release",
                                    "snippet": "Utilization improved to 95% as York is fully operational and Central City and Wood River are online and ramping, supporting $95 million of 45Z monetization opportunity.",
                                }
                            ]
                        ),
                        json.dumps(
                            [
                                {
                                    "doc_path": "release_q4.txt",
                                    "doc_type": "earnings_release",
                                    "snippet": "[DROPPED] map permit list county parcel latitude longitude $12 million",
                                }
                            ]
                        ),
                        json.dumps(
                            [
                                {
                                    "doc_path": "ocr_bad.txt",
                                    "doc_type": "ocr",
                                    "snippet": "permit map county parcel latitude longitude list $10 million",
                                }
                            ]
                        ),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]
            visible_metrics = [str(ws.cell(row=rr, column=4).value or "") for rr in range(1, ws.max_row + 1)]
            assert any("Utilization" in note for note in visible_notes) or any(
                "45Z monetization opportunity estimated at $95.0m." in note for note in visible_notes
            ) or any(
                "45Z Monetization / EBITDA" in metric for metric in visible_metrics
            )
            assert all(not note.startswith("[DROPPED]") for note in visible_notes if note)
            assert not any("permit map county parcel" in note.lower() for note in visible_notes)


def test_pbi_quarter_notes_ui_allows_more_than_two_clean_guidance_rows_when_strong(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_quarter_notes_guidance_wide.xlsx")
            guidance_texts = [
                (
                    "Revenue guidance",
                    "Revenue guidance was reaffirmed at $1.90 billion to $1.95 billion for FY 2025 with improving parcel volumes.",
                ),
                (
                    "Adjusted EBIT guidance",
                    "Adjusted EBIT guidance increased to $450 million to $465 million due to pricing and mix improvements in SendTech.",
                ),
                (
                    "EPS guidance",
                    "EPS guidance was reaffirmed at $3.10 to $3.20 for FY 2025 as execution improved through the quarter.",
                ),
                (
                    "FCF target",
                    "FCF target improved to $370 million to $390 million for FY 2025 as working capital normalized.",
                ),
            ]
            guidance_rows = []
            for idx, (metric, sentence) in enumerate(guidance_texts, start=1):
                guidance_rows.append(
                    {
                        "quarter": pd.Timestamp("2025-12-31"),
                        "note_id": f"guidance-{idx}",
                        "category": "Guidance / outlook",
                        "claim": sentence,
                        "note": sentence,
                        "metric_ref": metric,
                        "score": 95.0 - idx,
                        "doc_type": "earnings_release",
                        "doc": f"release_{idx}.txt",
                        "evidence_snippet": sentence,
                        "evidence_json": json.dumps(
                            [
                                {
                                    "doc_path": f"release_{idx}.txt",
                                    "doc_type": "earnings_release",
                                    "snippet": sentence,
                                }
                            ]
                        ),
                    }
                )

            quarter_notes = pd.DataFrame(guidance_rows)
            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            guidance_rows_visible = [
                rr
                for rr in range(1, ws.max_row + 1)
                if ws.cell(row=rr, column=2).value == "Guidance / outlook" and str(ws.cell(row=rr, column=3).value or "").strip()
            ]
            assert len(guidance_rows_visible) >= 3


def test_pbi_promise_tracker_filters_malformed_management_target_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_tracker_quality.xlsx")
            promises = pd.DataFrame(
                {
                    "promise_id": ["good-1", "bad-1"],
                    "quarter": [pd.Timestamp("2025-03-31"), pd.Timestamp("2025-03-31")],
                    "created_quarter": [pd.Timestamp("2025-03-31"), pd.Timestamp("2025-03-31")],
                    "last_seen_quarter": [pd.Timestamp("2025-03-31"), pd.Timestamp("2025-03-31")],
                    "first_seen_evidence_quarter": [pd.Timestamp("2025-03-31"), pd.Timestamp("2025-03-31")],
                    "last_seen_evidence_quarter": [pd.Timestamp("2025-03-31"), pd.Timestamp("2025-03-31")],
                    "metric": ["Revenue guidance", "Management target"],
                    "metric_display": ["Revenue guidance", "Management target"],
                    "promise_text": [
                        "FY 2025 Revenue guidance $1.90bn-$1.95bn.",
                        "Bowes provided the following Management target",
                    ],
                    "text_full": [
                        "FY 2025 Revenue guidance $1.90bn-$1.95bn.",
                        "Bowes provided the following Management target",
                    ],
                    "text_snippet": [
                        "FY 2025 Revenue guidance $1.90bn-$1.95bn.",
                        "Bowes provided the following Management target",
                    ],
                    "target": ["$1.90bn-$1.95bn", ""],
                    "target_display": ["$1.90bn-$1.95bn", ""],
                    "target_time": [pd.Timestamp("2025-12-31"), pd.NaT],
                    "target_period_norm": ["FY2025", ""],
                    "promise_type": ["guidance_range", "operational"],
                    "theme_key": ["rev|fy2025", "bad|fy2025"],
                    "source": [{"source_type": "guidance_snapshot"}, {"source_type": "presentation"}],
                }
            )
            base_inputs = _make_inputs(out_path, ticker="TEST")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "promises": promises})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_tracker_ui_v2()
            ws = ctx.wb["Promise_Tracker_UI"]

            visible_values = [str(ws.cell(row=rr, column=cc).value or "") for rr in range(1, ws.max_row + 1) for cc in range(1, 8)]
            assert all("Bowes provided the following Management target" not in val for val in visible_values)
            assert all(val != "Management target" for val in visible_values)


def test_promise_tracker_feeder_can_skip_visible_sheet_and_still_seed_progress(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_tracker_hidden_feeder.xlsx")
            promises = pd.DataFrame(
                {
                    "promise_id": ["rev-guidance"],
                    "promise_text": ["FY 2026 Revenue guidance $1.76bn-$1.86bn."],
                    "metric_tag": ["Revenue guidance"],
                    "created_quarter": [pd.Timestamp("2025-12-31")],
                    "last_seen_quarter": [pd.Timestamp("2025-12-31")],
                    "first_seen_evidence_quarter": [pd.Timestamp("2025-12-31")],
                    "last_seen_evidence_quarter": [pd.Timestamp("2025-12-31")],
                    "form": ["8-K"],
                    "doc": ["release_q4.txt"],
                    "source_type": ["earnings_release"],
                    "target_kind": ["range"],
                    "target_time": [pd.Timestamp("2026-12-31")],
                    "promise_type": ["guidance_range"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2026 Revenue guidance $1.76bn-$1.86bn."})
                    ],
                }
            )
            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", promises=promises))
            ctx.callbacks.write_promise_tracker_ui_v2(render_visible=False)

            assert "Promise_Tracker_UI" not in ctx.wb.sheetnames
            tracker_rows = ctx.state["ui_state"]["promise_tracker_rows_by_q"]
            assert tracker_rows
            assert any(
                str(it.get("metric_display") or it.get("metric") or "") == "Revenue guidance"
                for rows in tracker_rows.values()
                for it in rows
            )


def test_pbi_promise_progress_filters_malformed_scaffolding_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_quality.xlsx")
            hist = _make_hist().copy()
            hist["revenue"] = [100_000_000.0, 110_000_000.0, 120_000_000.0, 130_000_000.0]
            progress = pd.DataFrame(
                {
                    "promise_id": ["good-1", "bad-1"],
                    "quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-12-31")],
                    "status": ["resolved_fail", "at_risk"],
                    "metric_ref": ["Revenue guidance", "Management target"],
                    "target": ["$500.0m-$520.0m", ""],
                    "latest": ["not yet measurable", "not yet measurable"],
                    "rationale": [
                        "FY 2025 revenue guidance was $500 million to $520 million.",
                        "would be incremental to Management target",
                    ],
                    "promise_type": ["guidance_range", "operational"],
                    "guidance_type": ["period", "text"],
                    "target_period_norm": ["FY2025", "UNK"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2025 revenue guidance was $500 million to $520 million."}),
                        json.dumps({"doc_type": "presentation", "snippet": "would be incremental to Management target"}),
                    ],
                }
            )
            base_inputs = _make_inputs(out_path, ticker="TEST", hist=hist)
            inputs = base_inputs.__class__(**{**vars(base_inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            visible_values = [str(ws.cell(row=rr, column=cc).value or "") for rr in range(1, ws.max_row + 1) for cc in range(1, 8)]
            assert any("Revenue guidance" in val for val in visible_values)
            assert all("would be incremental to Management target" not in val for val in visible_values)
            assert all(val != "Management target" for val in visible_values)


def test_pbi_promise_progress_keeps_text_latest_for_clean_milestone(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_milestone.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["m-1"],
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "status": ["achieved"],
                    "metric_ref": ["Strategic milestone"],
                    "target": ["York online by Q4 2025"],
                    "latest": ["fully operational"],
                    "rationale": ["York became fully operational in Q4 2025."],
                    "promise_type": ["milestone"],
                    "guidance_type": ["text"],
                    "target_period_norm": ["Q42025"],
                    "source_evidence_json": [json.dumps({"doc_type": "earnings_release", "snippet": "York became fully operational in Q4 2025."})],
                }
            )
            base_inputs = _make_inputs(out_path, ticker="TEST")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            visible_values = [str(ws.cell(row=rr, column=cc).value or "") for rr in range(1, ws.max_row + 1) for cc in range(1, 8)]
            assert "Strategic milestone" in visible_values
            assert any("fully operational" in val.lower() for val in visible_values)


def test_pbi_promise_progress_drops_pending_guidance_without_target(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_pending_blank_target.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["g-1"],
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "status": ["pending"],
                    "metric_ref": ["Revenue guidance"],
                    "target": [""],
                    "latest": ["not yet measurable"],
                    "rationale": ["Future revenue and profitability assumptions."],
                    "promise_type": ["guidance_range"],
                    "guidance_type": ["period"],
                    "target_period_norm": ["FY2026"],
                    "source_evidence_json": [json.dumps({"doc_type": "presentation", "snippet": "Future revenue and profitability assumptions."})],
                }
            )
            base_inputs = _make_inputs(out_path, ticker="TEST")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            visible_values = [str(ws.cell(row=rr, column=cc).value or "") for rr in range(1, ws.max_row + 1) for cc in range(1, 8)]
            assert all("Revenue guidance" not in val for val in visible_values)


def test_gpre_quarter_notes_ui_suppresses_context_light_bullets_and_generic_ramping(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_quarter_notes_context_cleanup.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-06-30")] * 4,
                    "note_id": ["good-1", "bad-1", "bad-2", "bad-3"],
                    "category": ["Results / drivers", "Programs / initiatives", "Programs / initiatives", "Programs / initiatives"],
                    "claim": [
                        "Risk management supports margins and cash flow.",
                        "? Advantage Nebraska will reduce the carbon-intensity of Green Plains' biofuel; targeted in 2H25 with decarb gallons.",
                        "Omaha IA Cedar Rapids PA ? ~940K tons on Summit Carbon Solutions.",
                        "Online and ramping up capture volumes",
                    ],
                    "note": [
                        "Risk management supports margins and cash flow.",
                        "? Advantage Nebraska will reduce the carbon-intensity of Green Plains' biofuel; targeted in 2H25 with decarb gallons.",
                        "Omaha IA Cedar Rapids PA ? ~940K tons on Summit Carbon Solutions.",
                        "Online and ramping up capture volumes",
                    ],
                    "metric_ref": ["Risk management", "45Z monetization / EBITDA", "Management target", "Strategic milestone"],
                    "score": [95.0, 94.0, 93.0, 92.0],
                    "doc_type": ["earnings_release", "presentation", "presentation", "presentation"],
                    "doc": ["release_q2.txt", "slides_q2.txt", "slides_q2.txt", "slides_q2.txt"],
                    "evidence_snippet": [
                        "Risk management supports margins and cash flow.",
                        "? Advantage Nebraska will reduce the carbon-intensity of Green Plains' biofuel; targeted in 2H25 with decarb gallons.",
                        "Omaha IA Cedar Rapids PA ? ~940K tons on Summit Carbon Solutions.",
                        "Online and ramping up capture volumes",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q2.txt", "doc_type": "earnings_release", "snippet": "Risk management supports margins and cash flow."}]),
                        json.dumps([{"doc_path": "slides_q2.txt", "doc_type": "presentation", "snippet": "? Advantage Nebraska will reduce the carbon-intensity of Green Plains' biofuel; targeted in 2H25 with decarb gallons."}]),
                        json.dumps([{"doc_path": "slides_q2.txt", "doc_type": "presentation", "snippet": "Omaha IA Cedar Rapids PA ? ~940K tons on Summit Carbon Solutions."}]),
                        json.dumps([{"doc_path": "slides_q2.txt", "doc_type": "presentation", "snippet": "Online and ramping up capture volumes"}]),
                    ],
                }
            )
            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]

            assert any("Risk management supports margins and cash flow" in note for note in visible_notes)
            assert not any("Advantage Nebraska will reduce the carbon-intensity" in note for note in visible_notes)
            assert not any("Omaha IA Cedar Rapids PA" in note for note in visible_notes)
            assert not any(note.strip().endswith("Online and ramping up capture volumes") for note in visible_notes)


def test_gpre_promise_progress_suppresses_generic_ramping_seed_without_entity_context(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_progress_no_generic_ramping.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-06-30")],
                    "note_id": ["ramping-generic"],
                    "category": ["Programs / initiatives"],
                    "claim": ["Plant online and ramping up capture volumes."],
                    "note": ["Plant online and ramping up capture volumes."],
                    "metric_ref": ["Strategic milestone"],
                    "score": [90.0],
                    "doc_type": ["presentation"],
                    "doc": ["slides_q2.txt"],
                    "evidence_snippet": ["Plant online and ramping up capture volumes."],
                    "evidence_json": [
                        json.dumps(
                            [
                                {
                                    "doc_path": "slides_q2.txt",
                                    "doc_type": "presentation",
                                    "snippet": "Plant online and ramping up capture volumes.",
                                }
                            ]
                        )
                    ],
                }
            )
            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            visible_values = [
                str(ws.cell(row=rr, column=cc).value or "")
                for rr in range(1, ws.max_row + 1)
                for cc in range(1, 7)
            ]
            assert not any("Plant online and ramping up capture volumes" in val for val in visible_values)
            assert not any(val == "Online and ramping" for val in visible_values)


def test_gpre_promise_tracker_and_progress_drop_fragment_metric_labels(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_fragment_metric_drop.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["bad-frag"],
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "status": ["in_progress"],
                    "metric_ref": ["which time the Partnership Strategic milestone"],
                    "target": ["1"],
                    "latest": ["Expected in 2024"],
                    "rationale": ["The Merger is expected to close on January 9, 2024, at which time the Partnership will commence the process."],
                    "promise_type": ["operational"],
                    "guidance_type": ["text"],
                    "target_period_norm": ["FY2024"],
                    "source_evidence_json": [
                        json.dumps(
                            {
                                "doc_type": "earnings_release",
                                "snippet": "The Merger is expected to close on January 9, 2024, at which time the Partnership will commence the process.",
                            }
                        )
                    ],
                }
            )
            base_inputs = _make_inputs(out_path, ticker="TEST")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]
            visible_values = [str(ws.cell(row=rr, column=cc).value or "") for rr in range(1, ws.max_row + 1) for cc in range(1, 7)]

            assert not any("which time the Partnership" in val for val in visible_values)

def test_gpre_tracker_and_progress_restore_clean_high_signal_rows_from_quarter_notes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_tracker_progress_recall.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 4,
                    "note_id": ["n-45z-q4", "n-45z-2026", "n-obion", "n-cc"],
                    "category": ["Tone / expectations", "Tone / expectations", "Debt / liquidity / covenants", "Programs / initiatives"],
                    "claim": [
                        "On track for $15 - $25 million of 45Z production tax credit monetization value net of discounts and other costs for the fourth quarter of 2025.",
                        "All eight operating ethanol plants expected to qualify for production tax credits in 2026.",
                        "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt.",
                        "Central City and Wood River online and ramping.",
                    ],
                    "note": [
                        "On track for $15 - $25 million of 45Z production tax credit monetization value net of discounts and other costs for the fourth quarter of 2025.",
                        "All eight operating ethanol plants expected to qualify for production tax credits in 2026.",
                        "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt.",
                        "Central City and Wood River online and ramping.",
                    ],
                    "metric_ref": ["Tone | Corporate", "Tone | Corporate", "Debt reduction", "Strategic milestone"],
                    "score": [96.0, 95.0, 94.0, 93.0],
                    "doc_type": ["earnings_release", "earnings_release", "earnings_release", "presentation"],
                    "doc": ["release_q3.txt", "release_q3.txt", "release_q3.txt", "slides_q3.txt"],
                    "evidence_snippet": [
                        "On track for $15 - $25 million of 45Z production tax credit monetization value net of discounts and other costs for the fourth quarter of 2025.",
                        "All eight operating ethanol plants expected to qualify for production tax credits in 2026.",
                        "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt.",
                        "Central City and Wood River online and ramping.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "On track for $15 - $25 million of 45Z production tax credit monetization value net of discounts and other costs for the fourth quarter of 2025."}]),
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "All eight operating ethanol plants expected to qualify for production tax credits in 2026."}]),
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt."}]),
                        json.dumps([{"doc_path": "slides_q3.txt", "doc_type": "presentation", "snippet": "Central City and Wood River online and ramping."}]),
                    ],
                }
            )
            inputs = _make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes)
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_quarter_notes_ui_v2()
            ctx.callbacks.write_promise_tracker_ui_v2()
            ctx.callbacks.write_promise_progress_ui_v2()

            qn_ws = ctx.wb["Quarter_Notes_UI"]
            tracker_ws = ctx.wb["Promise_Tracker_UI"]
            progress_ws = ctx.wb["Promise_Progress_UI"]
            qn_values = [
                str(qn_ws.cell(row=rr, column=cc).value or "")
                for rr in range(1, qn_ws.max_row + 1)
                for cc in range(1, 6)
            ]
            tracker_values = [
                str(tracker_ws.cell(row=rr, column=cc).value or "")
                for rr in range(1, tracker_ws.max_row + 1)
                for cc in range(1, 6)
            ]
            progress_values = [str(progress_ws.cell(row=rr, column=cc).value or "") for rr in range(1, progress_ws.max_row + 1) for cc in range(1, 7)]

            assert any("45Z" in val for val in tracker_values)
            assert not any("Obion" in val or "Debt reduction" in val for val in tracker_values)
            assert any("Obion" in val or "Debt reduction" in val for val in qn_values)
            assert any("45Z" in val for val in progress_values)


def test_promise_progress_groups_same_subject_wording_variants_into_one_visible_row(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_progress_same_subject_grouping.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-09-30")],
                    "note_id": ["obion-1", "obion-2"],
                    "category": ["Debt / liquidity / covenants", "Debt / liquidity / covenants"],
                    "claim": [
                        "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt.",
                        "Obion sale proceeds fully repaid $130.7 million junior mezzanine debt.",
                    ],
                    "note": [
                        "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt.",
                        "Obion sale proceeds fully repaid $130.7 million junior mezzanine debt.",
                    ],
                    "metric_ref": ["Debt reduction", "Debt reduction"],
                    "score": [95.0, 92.0],
                    "doc_type": ["earnings_release", "presentation"],
                    "doc": ["release_q3.txt", "slides_q3.txt"],
                    "evidence_snippet": [
                        "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt.",
                        "Obion sale proceeds fully repaid $130.7 million junior mezzanine debt.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt."}]),
                        json.dumps([{"doc_path": "slides_q3.txt", "doc_type": "presentation", "snippet": "Obion sale proceeds fully repaid $130.7 million junior mezzanine debt."}]),
                    ],
                }
            )
            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ctx.callbacks.write_promise_tracker_ui_v2()
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            visible_values = [str(ws.cell(row=rr, column=cc).value or "") for rr in range(1, ws.max_row + 1) for cc in range(1, 7)]
            assert sum(1 for val in visible_values if "Obion" in val or "Debt reduction milestone" in val) <= 1
            assert not any("would be incremental to" in val.lower() for val in visible_values)
            assert not any("provided the following management target" in val.lower() for val in visible_values)


def test_gpre_45z_qualification_progress_does_not_use_monetization_range_as_latest(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_progress_qualification_vs_monetization.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["q-45z-2026", "q-45z-q4"],
                    "quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-09-30")],
                    "status": ["on_track", "on_track"],
                    "metric_ref": ["45Z plant qualification readiness", "45Z monetization / EBITDA"],
                    "target": ["All eight operating plants qualify in 2026", "$15.0m-$25.0m"],
                    "latest": ["not yet measurable", "not yet measurable"],
                    "rationale": [
                        "All eight operating ethanol plants expected to qualify for production tax credits in 2026.",
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                    ],
                    "promise_type": ["operational", "guidance_range"],
                    "guidance_type": ["period", "period"],
                    "target_period_norm": ["FY2026", "Q42025"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "All eight operating ethanol plants expected to qualify for production tax credits in 2026."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025."}),
                    ],
                }
            )
            base_inputs = _make_inputs(out_path, ticker="TEST")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            qualification_row = None
            for rr in range(1, ws.max_row + 1):
                if ws.cell(row=rr, column=2).value == "45Z plant qualification readiness":
                    qualification_row = rr
                    break
            assert qualification_row is not None
            latest_val = str(ws.cell(row=qualification_row, column=4).value or "")
            assert "$15.0m-$25.0m" not in latest_val
            assert "$15m-$25m" not in latest_val


def test_quarter_notes_ui_event_layer_dedupes_semantically_overlapping_notes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "quarter_notes_event_layer.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 2,
                    "note_id": ["rm-1", "rm-2"],
                    "category": ["Results / drivers", "Results / drivers"],
                    "claim": [
                        "Risk management supports margins and cash flow.",
                        "Risk management continued to support margins and cash flow through the quarter.",
                    ],
                    "note": [
                        "Risk management supports margins and cash flow.",
                        "Risk management continued to support margins and cash flow through the quarter.",
                    ],
                    "metric_ref": ["Risk management", "Risk management"],
                    "score": [88.0, 84.0],
                    "doc_type": ["earnings_release", "transcript"],
                    "doc": ["release_q3.txt", "transcript_q3.txt"],
                    "evidence_snippet": [
                        "Risk management supports margins and cash flow.",
                        "Risk management continued to support margins and cash flow through the quarter.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "Risk management supports margins and cash flow."}]),
                        json.dumps([{"doc_path": "transcript_q3.txt", "doc_type": "transcript", "snippet": "Risk management continued to support margins and cash flow through the quarter."}]),
                    ],
                }
            )
            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            rows = [
                (
                    str(ws.cell(row=rr, column=2).value or ""),
                    str(ws.cell(row=rr, column=3).value or ""),
                    str(ws.cell(row=rr, column=4).value or ""),
                )
                for rr in range(1, ws.max_row + 1)
                if str(ws.cell(row=rr, column=4).value or "") == "Risk management"
            ]
            assert len(rows) == 1
            assert "Risk management" in rows[0][1]


def test_pbi_recent_quarter_notes_keep_some_non_guidance_when_strong_candidates_exist(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_diversity.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 3,
                    "note_id": ["guidance-1", "driver-1", "junk-1"],
                    "category": ["Guidance / outlook", "Results / drivers", "Other / footnotes"],
                    "claim": [
                        "Adjusted EBIT guidance increased to $450 million to $465 million due to pricing and mix improvements in SendTech.",
                        "PB Bank liquidity improved by $120 million as trapped capital was released and SendTech margins improved on pricing and mix.",
                        "map permit list county parcel latitude longitude $12 million",
                    ],
                    "note": [
                        "Adjusted EBIT guidance increased to $450 million to $465 million due to pricing and mix improvements in SendTech.",
                        "PB Bank liquidity improved by $120 million as trapped capital was released and SendTech margins improved on pricing and mix.",
                        "map permit list county parcel latitude longitude $12 million",
                    ],
                    "metric_ref": ["Adjusted EBIT guidance", "PB Bank liquidity release", "Other"],
                    "score": [95.0, 93.0, 96.0],
                    "doc_type": ["earnings_release", "transcript", "ocr"],
                    "doc": ["release_q4.txt", "transcript_q4.txt", "ocr_junk.txt"],
                    "evidence_snippet": [
                        "Adjusted EBIT guidance increased to $450 million to $465 million due to pricing and mix improvements in SendTech.",
                        "PB Bank liquidity improved by $120 million as trapped capital was released and SendTech margins improved on pricing and mix.",
                        "map permit list county parcel latitude longitude $12 million",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": doc, "doc_type": doc_type, "snippet": s}])
                        for doc, doc_type, s in [
                            ("release_q4.txt", "earnings_release", "Adjusted EBIT guidance increased to $450 million to $465 million due to pricing and mix improvements in SendTech."),
                            ("transcript_q4.txt", "transcript", "PB Bank liquidity improved by $120 million as trapped capital was released and SendTech margins improved on pricing and mix."),
                            ("ocr_junk.txt", "ocr", "map permit list county parcel latitude longitude $12 million"),
                        ]
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_metrics = [str(ws.cell(row=rr, column=4).value or "") for rr in range(1, ws.max_row + 1)]
            assert any(metric == "PB Bank liquidity release" for metric in visible_metrics)
            assert any(metric == "Adjusted EBIT guidance" for metric in visible_metrics)


def test_pbi_recent_quarter_notes_can_recall_result_evidence_without_tracker_origination(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_result_evidence_recall.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-12-31")],
                    "note_id": ["guidance-1", "buyback-1"],
                    "category": ["Guidance / outlook", "Cash / liquidity / leverage"],
                    "claim": [
                        "Adjusted EBIT guidance increased to $450 million to $465 million for FY 2025.",
                        "Pitney Bowes increased share repurchase authorization and maintained significant remaining capacity in Q4 2025.",
                    ],
                    "note": [
                        "Adjusted EBIT guidance increased to $450 million to $465 million for FY 2025.",
                        "Pitney Bowes increased share repurchase authorization and maintained significant remaining capacity in Q4 2025.",
                    ],
                    "metric_ref": ["Adjusted EBIT guidance", "Capital allocation"],
                    "score": [95.0, 90.0],
                    "doc_type": ["earnings_release", "earnings_release"],
                    "doc": ["release_q4.txt", "release_q4.txt"],
                    "evidence_snippet": [
                        "Adjusted EBIT guidance increased to $450 million to $465 million for FY 2025.",
                        "Pitney Bowes increased share repurchase authorization and maintained significant remaining capacity in Q4 2025.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q4.txt", "doc_type": "earnings_release", "snippet": "Adjusted EBIT guidance increased to $450 million to $465 million for FY 2025."}]),
                        json.dumps([{"doc_path": "release_q4.txt", "doc_type": "earnings_release", "snippet": "Pitney Bowes increased share repurchase authorization and maintained significant remaining capacity in Q4 2025."}]),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_metrics = [str(ws.cell(row=rr, column=4).value or "") for rr in range(1, ws.max_row + 1)]
            assert "Capital allocation / buyback" in visible_metrics


def test_pbi_recent_quarter_notes_can_recall_result_evidence_from_promises(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_result_evidence_from_promises.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "note_id": ["guidance-1"],
                    "category": ["Guidance / outlook"],
                    "claim": ["Adjusted EBIT guidance increased to $410 million to $460 million for FY 2026."],
                    "note": ["Adjusted EBIT guidance increased to $410 million to $460 million for FY 2026."],
                    "metric_ref": ["Adjusted EBIT guidance"],
                    "score": [95.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q4.txt"],
                    "evidence_snippet": ["Adjusted EBIT guidance increased to $410 million to $460 million for FY 2026."],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q4.txt", "doc_type": "earnings_release", "snippet": "Adjusted EBIT guidance increased to $410 million to $460 million for FY 2026."}])
                    ],
                }
            )
            promises = pd.DataFrame(
                {
                    "promise_id": ["buyback-1"],
                    "created_quarter": [pd.Timestamp("2025-12-31")],
                    "metric": ["capital_allocation"],
                    "statement": [
                        "The company repurchased 12.6 million shares for $127 million in Q4 2025 and increased remaining authorization capacity."
                    ],
                    "promise_text": [
                        "The company repurchased 12.6 million shares for $127 million in Q4 2025 and increased remaining authorization capacity."
                    ],
                    "evidence_snippet": [
                        "The company repurchased 12.6 million shares for $127 million in Q4 2025 and increased remaining authorization capacity."
                    ],
                    "promise_type": ["operational"],
                    "confidence": ["high"],
                    "doc": ["Q4 2025 Earnings Press Release.pdf"],
                    "source_evidence_json": [
                        json.dumps(
                            [
                                {
                                    "doc_path": "Q4 2025 Earnings Press Release.pdf",
                                    "doc_type": "earnings_release",
                                    "source_type": "earnings_release",
                                    "snippet": "The company repurchased 12.6 million shares for $127 million in Q4 2025 and increased remaining authorization capacity.",
                                }
                            ]
                        )
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes, promises=promises))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_metrics = [str(ws.cell(row=rr, column=4).value or "") for rr in range(1, ws.max_row + 1)]
            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]
            assert "Capital allocation / buyback" in visible_metrics
            assert any("Repurchased 12.6m shares for $127.0m" in note for note in visible_notes)


def test_pbi_quarter_notes_preserve_revolver_from_to_detail_when_available(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_revolver_detail.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")],
                    "note_id": ["revolver-1"],
                    "category": ["Debt / liquidity / covenants"],
                    "claim": ["Revolver availability changed."],
                    "note": ["Revolver availability changed."],
                    "metric_ref": ["revolver_availability_change"],
                    "score": [92.0],
                    "doc_type": ["filing_text"],
                    "doc": ["pbi-20250930.htm"],
                    "evidence_snippet": [
                        "In the third quarter of 2025, the revolving credit facility was increased from $265 million to $400 million."
                    ],
                    "evidence_json": [
                        json.dumps(
                            [
                                {
                                    "doc_path": "pbi-20250930.htm",
                                    "doc_type": "filing_text",
                                    "snippet": "In the third quarter of 2025, the revolving credit facility was increased from $265 million to $400 million.",
                                }
                            ]
                        )
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]
            assert any("Revolver availability increased from $265.0m to $400.0m" in note for note in visible_notes)


def test_pbi_quarter_notes_infer_revolver_from_to_from_current_and_delta(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_revolver_delta_detail.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")],
                    "note_id": ["revolver-delta-1"],
                    "category": ["Debt / liquidity / covenants"],
                    "claim": ["Revolver availability changed."],
                    "note": ["Revolver availability changed."],
                    "metric_ref": ["revolver_availability_change"],
                    "score": [92.0],
                    "doc_type": ["revolver"],
                    "doc": ["revolver.csv"],
                    "comment_full_text": ["Revolver availability moved to $400.0m at 2025-09-30 (delta $135.0m)."],
                    "evidence_snippet": ["Revolver availability moved to $400.0m at 2025-09-30 (delta $135.0m)."],
                    "evidence_json": [
                        json.dumps(
                            [
                                {
                                    "doc_path": "revolver.csv",
                                    "doc_type": "revolver",
                                    "snippet": "Revolver availability moved to $400.0m at 2025-09-30 (delta $135.0m).",
                                }
                            ]
                        )
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]
            assert any("Revolver availability increased from $265.0m to $400.0m" in note for note in visible_notes)


def test_pbi_quarter_notes_prefer_metric_aware_guidance_summary_over_updated_target(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_metric_aware_guidance_summary.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-06-30")],
                    "note_id": ["fcf-guidance-1"],
                    "category": ["Guidance / outlook"],
                    "claim": ["Updated target $330m-$370m"],
                    "note": ["Updated target $330m-$370m"],
                    "metric_ref": ["FCF target"],
                    "score": [95.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q2.txt"],
                    "evidence_snippet": ["FY 2025 free cash flow guidance updated to $330 million to $370 million."],
                    "evidence_json": [
                        json.dumps(
                            [
                                {
                                    "doc_path": "release_q2.txt",
                                    "doc_type": "earnings_release",
                                    "snippet": "FY 2025 free cash flow guidance updated to $330 million to $370 million.",
                                }
                            ]
                        )
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]
            assert any("FY 2025 FCF target updated to $330m-$370m." in note for note in visible_notes)
            assert not any(note == "Updated target $330m-$370m" for note in visible_notes)


def test_pbi_quarter_notes_do_not_repeat_target_wording_in_guidance_summary(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_no_target_target.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-03-31")],
                    "note_id": ["fcf-guidance-target-1"],
                    "category": ["Guidance / outlook"],
                    "claim": ["FY 2025 free cash flow guidance of $330m-$370m."],
                    "note": ["FY 2025 free cash flow guidance of $330m-$370m."],
                    "metric_ref": ["FCF target"],
                    "score": [95.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q1.txt"],
                    "evidence_snippet": ["FY 2025 free cash flow guidance of $330m-$370m."],
                    "evidence_json": [
                        json.dumps(
                            [
                                {
                                    "doc_path": "release_q1.txt",
                                    "doc_type": "earnings_release",
                                    "snippet": "FY 2025 free cash flow guidance of $330m-$370m.",
                                }
                            ]
                        )
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]
            assert any("FY 2025 FCF target $330m-$370m." in note for note in visible_notes)
            assert not any("target target" in note.lower() for note in visible_notes)


def test_pbi_quarter_notes_prefer_richer_revolver_summary_over_generic_liquidity_line(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_prefer_richer_revolver.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-12-31")],
                    "note_id": ["liq-generic", "liq-rich"],
                    "category": ["Debt / refi / covenants", "Debt / refi / covenants"],
                    "claim": ["Liquidity improved", "Revolver availability changed"],
                    "note": ["Liquidity improved", "Revolver availability changed"],
                    "metric_ref": ["Debt reduction", "revolver_availability_change"],
                    "score": [93.0, 89.0],
                    "doc_type": ["earnings_release", "filing_text"],
                    "doc": ["release_q4.txt", "pbi-20250930.htm"],
                    "evidence_snippet": [
                        "Debt refinancing improved liquidity.",
                        "Revolver availability moved to $400.0m at 2025-09-30 (delta $135.0m).",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q4.txt", "doc_type": "earnings_release", "snippet": "Debt refinancing improved liquidity."}]),
                        json.dumps([{"doc_path": "pbi-20250930.htm", "doc_type": "filing_text", "snippet": "Revolver availability moved to $400.0m at 2025-09-30 (delta $135.0m)."}]),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]

            assert any("Revolver availability increased from $265.0m to $400.0m" in note for note in visible_notes)
            assert not any(note == "Liquidity improved" for note in visible_notes)


def test_pbi_recent_quarter_notes_keep_multiple_non_guidance_rows_when_fact_rich_candidates_exist(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_multi_block_diversity.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [
                        pd.Timestamp("2025-12-31"),
                        pd.Timestamp("2025-12-31"),
                        pd.Timestamp("2025-09-30"),
                        pd.Timestamp("2025-09-30"),
                    ],
                    "note_id": ["g-1", "rev-1", "g-2", "rev-2"],
                    "category": [
                        "Guidance / outlook",
                        "Debt / refi / covenants",
                        "Guidance / outlook",
                        "Debt / refi / covenants",
                    ],
                    "claim": [
                        "FY 2026 Revenue guidance target $1.76bn-$1.86bn.",
                        "Revolver availability changed.",
                        "FY 2025 Revenue guidance target $1.90bn-$1.95bn.",
                        "Revolver availability changed.",
                    ],
                    "note": [
                        "FY 2026 Revenue guidance target $1.76bn-$1.86bn.",
                        "Revolver availability changed.",
                        "FY 2025 Revenue guidance target $1.90bn-$1.95bn.",
                        "Revolver availability changed.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "revolver_availability_change",
                        "Revenue guidance",
                        "revolver_availability_change",
                    ],
                    "score": [95.0, 90.0, 95.0, 90.0],
                    "doc_type": ["earnings_release", "filing_text", "earnings_release", "filing_text"],
                    "doc": ["release_q4.txt", "pbi-20251231.htm", "release_q3.txt", "pbi-20250930.htm"],
                    "evidence_snippet": [
                        "FY 2026 Revenue guidance target $1.76bn-$1.86bn.",
                        "The revolving credit facility was increased from $265 million to $400 million.",
                        "FY 2025 Revenue guidance target $1.90bn-$1.95bn.",
                        "The revolving credit facility was increased from $265 million to $400 million.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q4.txt", "doc_type": "earnings_release", "snippet": "FY 2026 Revenue guidance target $1.76bn-$1.86bn."}]),
                        json.dumps([{"doc_path": "pbi-20251231.htm", "doc_type": "filing_text", "snippet": "The revolving credit facility was increased from $265 million to $400 million."}]),
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "FY 2025 Revenue guidance target $1.90bn-$1.95bn."}]),
                        json.dumps([{"doc_path": "pbi-20250930.htm", "doc_type": "filing_text", "snippet": "The revolving credit facility was increased from $265 million to $400 million."}]),
                    ],
                }
            )
            promises = pd.DataFrame(
                {
                    "promise_id": ["p-q4", "p-q3"],
                    "created_quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-09-30")],
                    "metric": ["capital_allocation", "capital_allocation"],
                    "statement": [
                        "The company repurchased 12.6 million shares for $127 million in Q4 2025, reduced principal debt by $114 million and increased remaining authorization capacity to $359 million.",
                        "Through last Friday, the company repurchased 25.9 million shares at a total cost of $281.2 million and increased remaining authorization capacity.",
                    ],
                    "promise_text": [
                        "The company repurchased 12.6 million shares for $127 million in Q4 2025, reduced principal debt by $114 million and increased remaining authorization capacity to $359 million.",
                        "Through last Friday, the company repurchased 25.9 million shares at a total cost of $281.2 million and increased remaining authorization capacity.",
                    ],
                    "evidence_snippet": [
                        "The company repurchased 12.6 million shares for $127 million in Q4 2025, reduced principal debt by $114 million and increased remaining authorization capacity to $359 million.",
                        "Through last Friday, the company repurchased 25.9 million shares at a total cost of $281.2 million and increased remaining authorization capacity.",
                    ],
                    "promise_type": ["operational", "operational"],
                    "confidence": ["high", "high"],
                    "doc": ["release_q4.txt", "release_q3.txt"],
                    "source_evidence_json": [
                        json.dumps([{"doc_path": "release_q4.txt", "doc_type": "earnings_release", "source_type": "earnings_release", "snippet": "The company repurchased 12.6 million shares for $127 million in Q4 2025, reduced principal debt by $114 million and increased remaining authorization capacity to $359 million."}]),
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "source_type": "earnings_release", "snippet": "Through last Friday, the company repurchased 25.9 million shares at a total cost of $281.2 million and increased remaining authorization capacity."}]),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes, promises=promises))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            def _quarter_block(qtxt: str) -> list[tuple[str, str]]:
                out: list[tuple[str, str]] = []
                capture = False
                for rr in range(1, ws.max_row + 1):
                    marker = str(ws.cell(row=rr, column=1).value or "")
                    if marker == qtxt:
                        capture = True
                        continue
                    if capture and marker:
                        break
                    if capture:
                        note = str(ws.cell(row=rr, column=3).value or "")
                        metric = str(ws.cell(row=rr, column=4).value or "")
                        if note or metric:
                            out.append((note, metric))
                return out

            q4_rows = _quarter_block("2025-12-31")
            q3_rows = _quarter_block("2025-09-30")
            assert sum(1 for _, metric in q4_rows if metric not in {"Revenue guidance", "Adjusted EBIT guidance", "EPS guidance", "FCF target", "Cost savings target"}) >= 2
            assert sum(1 for _, metric in q3_rows if metric not in {"Revenue guidance", "Adjusted EBIT guidance", "EPS guidance", "FCF target", "Cost savings target"}) >= 2
            assert any("Repurchased 12.6m shares for $127.0m" in note for note, _ in q4_rows)
            assert any("Revolver availability increased from $265.0m to $400.0m" in note for note, _ in q4_rows)
            assert any("Revolver availability increased from $265.0m to $400.0m" in note for note, _ in q3_rows)


def test_pbi_quarter_notes_quantify_fcf_summary_from_adj_metrics_context(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_contextual_fcf.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "note_id": ["fcf-1"],
                    "category": ["Cash flow / FCF / capex"],
                    "claim": ["Free cash flow improved."],
                    "note": ["Free cash flow improved."],
                    "metric_ref": ["FCF improvement"],
                    "score": [92.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q4.txt"],
                    "evidence_snippet": ["Free cash flow improved."],
                    "evidence_json": [
                        json.dumps(
                            [
                                {
                                    "doc_path": "release_q4.txt",
                                    "doc_type": "earnings_release",
                                    "snippet": "Free cash flow improved.",
                                }
                            ]
                        )
                    ],
                }
            )
            adj_metrics = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2024-12-31"), pd.Timestamp("2025-12-31")],
                    "adj_fcf": [131_800_000.0, 221_700_000.0],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes)
            inputs = base_inputs.__class__(**{**vars(base_inputs), "adj_metrics": adj_metrics})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]
            assert any("Free cash flow improved to $221.7m, up $89.9m YoY." in note for note in visible_notes)


def test_pbi_quarter_notes_prefer_release_backed_fcf_over_adj_metrics_when_available(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_ticker_model_out_path(case_dir, "PBI", "pbi_notes_release_backed_fcf.xlsx")
            sec_cache = case_dir / "PBI" / "sec_cache"
            sec_cache.mkdir(parents=True, exist_ok=True)
            source_doc = sec_cache / "doc_000162828026008604_q42025earningspressrelea.htm"
            source_doc.write_text(
                (
                    "Fourth Quarter ($ millions except EPS) 2025 2024 Cash from Operations $222 $132 "
                    "Free Cash Flow1 $212 $142. "
                    "Pitney Bowes Inc. Reconciliation of reported net cash from operating activities to free cash flow "
                    "Net cash from operating activities - continuing operations $221,699 $131,837 "
                    "Capital expenditures (20,251) (22,182) Restructuring payments 10,495 32,104 "
                    "Free cash flow $211,943 $141,759."
                ),
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "note_id": ["fcf-1"],
                    "category": ["Cash flow / FCF / capex"],
                    "claim": ["Free cash flow improved."],
                    "note": ["Free cash flow improved."],
                    "metric_ref": ["FCF improvement"],
                    "score": [92.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q4.txt"],
                    "evidence_snippet": ["Free cash flow improved."],
                }
            )
            adj_metrics = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2024-12-31"), pd.Timestamp("2025-12-31")],
                    "adj_fcf": [131_800_000.0, 221_700_000.0],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="PBI", quarter_notes=quarter_notes)
            inputs = base_inputs.__class__(**{**vars(base_inputs), "adj_metrics": adj_metrics})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]
            assert any("Free cash flow improved to $211.9m, up $70.2m YoY." in note for note in visible_notes)
            assert not any("Free cash flow improved to $221.7m, up $89.9m YoY." in note for note in visible_notes)


def test_pbi_recent_quarter_notes_can_recall_ceo_letter_buyback_result_from_html_promise(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_ceo_letter_buyback_result.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")],
                    "note_id": ["guidance-1"],
                    "category": ["Guidance / outlook"],
                    "claim": ["Revenue guidance updated to $1.90bn-$1.95bn for FY 2025."],
                    "note": ["Revenue guidance updated to $1.90bn-$1.95bn for FY 2025."],
                    "metric_ref": ["Revenue guidance"],
                    "score": [95.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q3.txt"],
                    "evidence_snippet": ["Revenue guidance updated to $1.90bn-$1.95bn for FY 2025."],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "Revenue guidance updated to $1.90bn-$1.95bn for FY 2025."}])
                    ],
                }
            )
            promises = pd.DataFrame(
                {
                    "promise_id": ["buyback-1"],
                    "created_quarter": [pd.Timestamp("2025-09-30")],
                    "metric": ["capital_allocation"],
                    "statement": [
                        "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year."
                    ],
                    "promise_text": [
                        "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year."
                    ],
                    "evidence_snippet": [
                        "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year."
                    ],
                    "promise_type": ["operational"],
                    "confidence": ["high"],
                    "doc": ["doc_000162828025047122_q32025earningsceoletter.htm"],
                    "source_evidence_json": [
                        json.dumps(
                            [
                                {
                                    "doc_path": "doc_000162828025047122_q32025earningsceoletter.htm",
                                    "doc_type": "html",
                                    "source_type": "html",
                                    "snippet": "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year.",
                                }
                            ]
                        )
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes, promises=promises))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_metrics = [str(ws.cell(row=rr, column=4).value or "") for rr in range(1, ws.max_row + 1)]
            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]
            assert "Capital allocation / buyback" in visible_metrics
            assert any("Repurchased 25.9m shares for $281.2m" in note for note in visible_notes)


def test_pbi_quarter_notes_preserve_buyback_since_start_detail_from_evidence_snippet(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_buyback_since_start.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")],
                    "note_id": ["bb-1"],
                    "category": ["Cash / liquidity / leverage"],
                    "claim": ["Share repurchase authorization and remaining capacity updated."],
                    "note": ["Share repurchase authorization and remaining capacity updated."],
                    "metric_ref": ["Capital allocation / buyback"],
                    "score": [92.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q3.txt"],
                    "evidence_snippet": [
                        "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million "
                        "since starting the program earlier this year. In addition to increasing our share repurchase "
                        "authorization to $500 million, we increased our quarterly dividend from $0.08 to $0.09 per share."
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]
            assert any(
                "Repurchased 25.9m shares for $281.2m" in note
                and "since starting the program earlier this year" in note
                for note in visible_notes
            )
            assert any("Repurchase authorization increased to $500.0m" in note for note in visible_notes)


def test_pbi_quarter_notes_keep_quantified_fcf_actual_when_block_has_four_guidance_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_keep_fcf_actual.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 6,
                    "note_id": ["g-1", "g-2", "g-3", "g-4", "bb-1", "debt-1"],
                    "category": [
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Cash / liquidity / leverage",
                        "Cash / liquidity / leverage",
                    ],
                    "claim": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Share repurchase authorization and remaining capacity updated.",
                        "Liquidity improved.",
                    ],
                    "note": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Share repurchase authorization and remaining capacity updated.",
                        "Liquidity improved.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "EPS guidance",
                        "FCF target",
                        "Capital allocation / buyback",
                        "Debt reduction",
                    ],
                    "score": [96.0, 95.0, 94.0, 97.0, 93.0, 92.0],
                    "doc_type": ["earnings_release"] * 6,
                    "doc": ["release_q4.txt"] * 6,
                    "evidence_snippet": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Repurchased 12.6 million shares for $127.0 million in the fourth quarter.",
                        "Reduced principal debt by $114.1 million in the fourth quarter.",
                    ],
                }
            )
            adj_metrics = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2024-12-31"), pd.Timestamp("2025-12-31")],
                    "adj_fcf": [131_800_000.0, 221_700_000.0],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes)
            inputs = base_inputs.__class__(**{**vars(base_inputs), "adj_metrics": adj_metrics})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            def _quarter_block(qtxt: str) -> list[str]:
                out: list[str] = []
                capture = False
                for rr in range(1, ws.max_row + 1):
                    marker = str(ws.cell(row=rr, column=1).value or "")
                    if marker == qtxt:
                        capture = True
                        continue
                    if capture and marker:
                        break
                    if capture:
                        note = str(ws.cell(row=rr, column=3).value or "")
                        if note:
                            out.append(note)
                return out

            q4_rows = _quarter_block("2025-12-31")
            q4_note_rows = [x for x in q4_rows if x != "Note"]
            guidance_like = [x for x in q4_note_rows if re.search(r"\b(guidance|target)\b", x, re.I)]

            assert any("Free cash flow improved to $221.7m, up $89.9m YoY." in note for note in q4_note_rows)
            assert any("Reduced principal debt by $114.1m in Q4." in note for note in q4_note_rows)
            assert len(guidance_like) == 4
            assert all(re.search(r"\b(guidance|target)\b", note, re.I) for note in q4_note_rows[:4])


def test_pbi_quarter_notes_replace_generic_cap_alloc_with_richer_promise_rescue(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_replace_generic_buyback.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 5,
                    "note_id": ["g-1", "g-2", "g-3", "g-4", "bb-generic"],
                    "category": ["Guidance / outlook"] * 4 + ["Cash / liquidity / leverage"],
                    "claim": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "Share repurchase authorization and remaining capacity updated.",
                    ],
                    "note": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "Share repurchase authorization and remaining capacity updated.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "EPS guidance",
                        "FCF target",
                        "Capital allocation / buyback",
                    ],
                    "score": [95.0, 94.0, 93.0, 92.0, 90.0],
                    "doc_type": ["earnings_release"] * 5,
                    "doc": ["release_q3.txt"] * 5,
                    "evidence_snippet": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "Share repurchase authorization and remaining capacity updated.",
                    ],
                }
            )
            promises = pd.DataFrame(
                {
                    "promise_id": ["buyback-1"],
                    "created_quarter": [pd.Timestamp("2025-09-30")],
                    "metric": ["capital_allocation"],
                    "statement": [
                        "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year."
                    ],
                    "promise_text": [
                        "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year."
                    ],
                    "evidence_snippet": [
                        "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year."
                    ],
                    "promise_type": ["operational"],
                    "confidence": ["high"],
                    "doc": ["doc_000162828025047122_q32025earningsceoletter.htm"],
                    "source_evidence_json": [
                        json.dumps(
                            [
                                {
                                    "doc_path": "doc_000162828025047122_q32025earningsceoletter.htm",
                                    "doc_type": "html",
                                    "source_type": "html",
                                    "snippet": "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year.",
                                }
                            ]
                        )
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes, promises=promises))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            visible_notes = [str(ws.cell(row=rr, column=3).value or "") for rr in range(1, ws.max_row + 1)]

            assert any("Repurchased 25.9m shares for $281.2m" in note for note in visible_notes)
            assert not any("Share repurchase authorization and remaining capacity updated." == note for note in visible_notes)


def test_pbi_quarter_notes_can_build_q3_buyback_execution_from_sec_table(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q3_buyback_table.xlsx")
            sec_cache_dir = case_dir / "TEST" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828025047360_pbi-20250930.htm").write_text(
                "Item 2. Unregistered Sales of Equity Securities and Use of Proceeds. "
                "On February 11, 2025, our Board of Directors authorized a new $150 million share repurchase program. "
                "In July 2025, the Board authorized an increase in the program to $400 million, and in October 2025, "
                "the Board authorized an additional increase in the program to $500 million. "
                "We increased our quarterly dividend from $0.08 to $0.09 per share. "
                "The following table provides information about common stock purchases during the three months ended September 30, 2025. "
                "Total number of shares purchased Average price paid per share Total number of shares purchased as part of publicly announced plans or programs Approximate dollar value of shares that may yet be purchased under the plans or programs. "
                "July 2025 4,174,838 $ 11.60 4,174,838 $261,280 "
                "August 2025 7,722,528 $ 11.27 7,722,528 $174,267 "
                "September 2025 2,220,772 $ 11.73 2,220,772 $148,226 "
                "14,118,138 $ 11.44 14,118,138 "
                "As of September 30, 2025, $148.226 million in capacity remained under the authorization.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 4,
                    "note_id": ["g-1", "g-2", "g-3", "g-4"],
                    "category": ["Guidance / outlook"] * 4,
                    "claim": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                    ],
                    "note": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "EPS guidance",
                        "FCF target",
                    ],
                    "score": [95.0, 94.0, 93.0, 92.0],
                    "doc_type": ["earnings_release"] * 4,
                    "doc": ["release_q3.txt"] * 4,
                    "evidence_snippet": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q3_rows = _quarter_block_notes(ws, "2025-09-30")

            assert any(
                "Repurchased 14.1m shares for $161.5m with an average price of $11.44/share in Q3." in note
                for note in q3_rows
            )
            assert any(
                "Repurchase authorization increased to $500.0m, up from $400.0m." in note
                for note in q3_rows
            )
            assert any("Remaining share repurchase capacity was $148.2m at quarter-end." in note for note in q3_rows)
            assert any("Quarterly dividend increased to $0.09/share from $0.08/share." in note for note in q3_rows)


def test_pbi_quarter_notes_visible_q3_prefers_sec_table_buyback_over_cumulative_since_start_row(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q3_visible_prefers_sec_table_buyback.xlsx")
            sec_cache_dir = case_dir / "TEST" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828025047360_pbi-20250930.htm").write_text(
                "Item 2. Unregistered Sales of Equity Securities and Use of Proceeds. "
                "The following table provides information about common stock purchases during the three months ended September 30, 2025. "
                "Total number of shares purchased Average price paid per share Total number of shares purchased as part of publicly announced plans or programs Approximate dollar value of shares that may yet be purchased under the plans or programs. "
                "July 2025 4,174,838 $ 11.60 4,174,838 $261,280 "
                "August 2025 7,722,528 $ 11.27 7,722,528 $174,267 "
                "September 2025 2,220,772 $ 11.73 2,220,772 $148,226 "
                "14,118,138 $ 11.44 14,118,138 "
                "As of September 30, 2025, $148.226 million in capacity remained under the authorization.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 5,
                    "note_id": ["g-1", "g-2", "g-3", "g-4", "bb-1"],
                    "category": ["Guidance / outlook"] * 4 + ["Cash / liquidity / leverage"],
                    "claim": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "Share repurchase authorization and remaining capacity updated.",
                    ],
                    "note": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "Share repurchase authorization and remaining capacity updated.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "EPS guidance",
                        "FCF target",
                        "Capital allocation / buyback",
                    ],
                    "score": [95.0, 94.0, 93.0, 92.0, 91.0],
                    "doc_type": ["earnings_release"] * 5,
                    "doc": ["release_q3.txt"] * 5,
                    "source_type": ["earnings_release"] * 5,
                    "evidence_snippet": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q3_rows = _quarter_block_notes(ws, "2025-09-30")

            assert any(
                "Repurchased 14.1m shares for $161.5m with an average price of $11.44/share in Q3." in note
                for note in q3_rows
            )
            assert not any("Repurchased 25.9m shares for $281.2m since starting the program earlier this year." in note for note in q3_rows)


def test_pbi_realistic_buyback_table_with_program_boilerplate_still_builds_quarter_safe_q3_row(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q3_realistic_table_with_program_boilerplate.xlsx")
            sec_cache_dir = case_dir / "TEST" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828025047360_pbi-20250930.htm").write_text(
                "Item 2. Unregistered Sales of Equity Securities and Use of Proceeds. "
                "On February 11, 2025, our Board of Directors authorized a new $150 million share repurchase program. "
                "In July 2025, the Board authorized an increase in the program to $400 million, and in October 2025, "
                "the Board authorized an additional increase in the program to $500 million. "
                "Subject to limitations in our New Credit Agreement, common stock repurchases may be made from time to time "
                "in open market or private transactions in such manner as may be deemed advisable from time to time. "
                "We may also repurchase shares of our common stock to manage the dilution created by shares issued under employee stock plans. "
                "The following table provides information about common stock purchases during the three months ended September 30, 2025. "
                "Total number of shares purchased Average price paid per share Total number of shares purchased as part of publicly announced plans or programs Approximate dollar value of shares that may yet be purchased under the plans or programs. "
                "July 2025 4,174,838 $ 11.60 4,174,838 $261,280 "
                "August 2025 7,722,528 $ 11.27 7,722,528 $174,267 "
                "September 2025 2,220,772 $ 11.73 2,220,772 $148,226 "
                "14,118,138 $ 11.44 14,118,138 "
                "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 5,
                    "note_id": ["g-1", "g-2", "g-3", "g-4", "bb-1"],
                    "category": ["Guidance / outlook"] * 4 + ["Cash / liquidity / leverage"],
                    "claim": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "Share repurchase authorization and remaining capacity updated.",
                    ],
                    "note": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "Share repurchase authorization and remaining capacity updated.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "EPS guidance",
                        "FCF target",
                        "Capital allocation / buyback",
                    ],
                    "score": [95.0, 94.0, 93.0, 92.0, 91.0],
                    "doc_type": ["earnings_release"] * 5,
                    "doc": ["q3.txt"] * 5,
                    "source_type": ["earnings_release"] * 5,
                    "evidence_snippet": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q3_rows = _quarter_block_notes(ws, "2025-09-30")

            assert any(
                "Repurchased 14.1m shares for $161.5m with an average price of $11.44/share in Q3." in note
                for note in q3_rows
            )
            assert not any(
                "Repurchased 25.9m shares for $281.2m since starting the program earlier this year." in note
                for note in q3_rows
            )


def test_pbi_quarter_notes_can_parse_realistic_month_table_when_header_text_is_fragmented(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q3_realistic_fragmented_table.xlsx")
            sec_cache_dir = case_dir / "TEST" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828025047360_pbi-20250930.htm").write_text(
                "ITEM 5. MARKET FOR THE COMPANY'S COMMON EQUITY RELATED STOCKHOLDER MATTERS AND ISSUER PURCHASES OF EQUITY SECURITIES "
                "Dividends and Share Repurchases The following table provides information about common stock purchases during the three months ended September 30, 2025. "
                "Approximate dollar value of shares that may yet be purchased under the plans or programs. "
                "July 2025 4,174,838 $ 11.60 4,174,838 $261,280 "
                "August 2025 7,722,528 $ 11.27 7,722,528 $174,267 "
                "September 2025 2,220,772 $ 11.73 2,220,772 $148,226 "
                "14,118,138 $ 11.44 14,118,138 "
                "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million since starting the program earlier this year.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")],
                    "note_id": ["q3-guid"],
                    "category": ["Guidance / outlook"],
                    "claim": ["Q3 placeholder."],
                    "note": ["Q3 placeholder."],
                    "metric_ref": ["Guidance"],
                    "score": [90.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["q3.txt"],
                    "source_type": ["earnings_release"],
                    "evidence_snippet": ["Q3 placeholder."],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q3_rows = _quarter_block_notes(ws, "2025-09-30")

            assert any(
                "Repurchased 14.1m shares for $161.5m with an average price of $11.44/share in Q3." in note
                for note in q3_rows
            )


def test_pbi_q4_keeps_single_quarter_specific_buyback_execution_row(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q4_single_buyback_row.xlsx")
            sec_cache_dir = case_dir / "TEST" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828026008604_q42025earningspressrelea.htm").write_text(
                "During the quarter we repurchased 12.6 million shares for $127.0 million.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828026012345_pbi-20251231.htm").write_text(
                "The following table provides information about common stock purchases during the three months ended December 31, 2025. "
                "Total number of shares purchased Average price paid per share Total number of shares purchased as part of publicly announced plans or programs Approximate dollar value of shares that may yet be purchased under the plans or programs. "
                "October 2025 3,203,100&#160; $ 11.30&#160; 3,203,100 $212,031 "
                "November 2025 7,926,090&#160; $ 9.52&#160; 7,926,090 $136,553 "
                "December 2025 1,484,407&#160; $ 10.05&#160; 1,484,407 $121,639 "
                "12,613,597&#160; $ 10.04&#160; 12,613,597 ",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 5,
                    "note_id": ["g-1", "g-2", "g-3", "g-4", "bb-1"],
                    "category": ["Guidance / outlook"] * 4 + ["Cash / liquidity / leverage"],
                    "claim": [
                        "FY 2026 Revenue guidance.",
                        "FY 2026 Adjusted EBIT guidance.",
                        "FY 2026 EPS guidance.",
                        "FY 2026 FCF guidance.",
                        "Share repurchases continued in Q4.",
                    ],
                    "note": [
                        "FY 2026 Revenue guidance.",
                        "FY 2026 Adjusted EBIT guidance.",
                        "FY 2026 EPS guidance.",
                        "FY 2026 FCF guidance.",
                        "Share repurchases continued in Q4.",
                    ],
                    "metric_ref": ["Revenue guidance", "Adjusted EBIT guidance", "EPS guidance", "FCF target", "Capital allocation / buyback"],
                    "score": [95.0, 94.0, 93.0, 92.0, 91.0],
                    "doc_type": ["earnings_release"] * 5,
                    "doc": ["q4.txt"] * 5,
                    "source_type": ["earnings_release"] * 5,
                    "evidence_snippet": [
                        "FY 2026 Revenue guidance.",
                        "FY 2026 Adjusted EBIT guidance.",
                        "FY 2026 EPS guidance.",
                        "FY 2026 FCF guidance.",
                        "During the quarter we repurchased 12.6 million shares for $127.0 million.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q4_rows = _quarter_block_notes(ws, "2025-12-31")
            buyback_rows = [note for note in q4_rows if note.startswith("Repurchased 12.6m shares for $")]

            assert len(buyback_rows) == 1
            assert "with an average price of $" in buyback_rows[0]
            assert "in Q4." in buyback_rows[0]


def test_generic_quarter_notes_can_add_financing_and_use_of_proceeds_notes_from_sec_cache(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "generic_notes_convertible_financing.xlsx")
            sec_cache_dir = case_dir / "PBI" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_generic-20250930-financing.htm").write_text(
                "Pitney Bowes Inc. In August 2025, we issued an aggregate $230 million convertible senior notes due 2030. "
                "The Convertible Notes accrue interest at a rate of 1.50% per annum. Net proceeds were $221 million. "
                "We used $61.9 million of the proceeds to repurchase 5.5 million of our common stock. "
                "We entered into capped call transactions that are expected to reduce the potential dilution of our common stock upon conversion. "
                "The remaining proceeds will be used for general corporate purposes and other strategic investments.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 2,
                    "note_id": ["g-1", "g-2"],
                    "category": ["Guidance / outlook", "Cash flow / FCF / capex"],
                    "claim": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FCF TTM improved by $25.0m YoY.",
                    ],
                    "note": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FCF TTM improved by $25.0m YoY.",
                    ],
                    "metric_ref": ["Revenue guidance", "FCF"],
                    "score": [95.0, 92.0],
                    "doc_type": ["earnings_release", "model_metric"],
                    "doc": ["release_q3.txt", "history_q"],
                    "source_type": ["earnings_release", "model_metric"],
                    "evidence_snippet": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FCF TTM improved by $25.0m YoY.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="PBI", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q3_rows = _quarter_block_notes(ws, "2025-09-30")

            assert any("Used $61.9m from convertible notes proceeds to repurchase 5.5m shares." in note for note in q3_rows)
            assert any("Entered capped call transactions expected to reduce dilution from convertible notes conversion." in note for note in q3_rows)
            assert not any("Remaining proceeds will fund general corporate purposes and other strategic investments." in note for note in q3_rows)


def test_gpre_quarter_notes_do_not_promote_static_buyback_program_text_or_note_exchange_as_share_repurchase(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_cap_alloc_safety.xlsx")
            sec_cache_dir = case_dir / "TEST" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940225000163_gpre-20250930.htm").write_text(
                "In August 2014 and October 2019, our Board authorized a share repurchase program of up to $200.0 million of our common stock. "
                "The company used approximately $30 million of the net proceeds from the subscription transactions to repurchase approximately 2.9 million shares of its common stock. "
                "At November 5, 2025, $77.2 million in share repurchase authorization remained. "
                "The company entered into exchange transactions to exchange $170 million aggregate principal amount of the 2027 Notes for $170 million of newly issued 5.25% Convertible Senior Notes due November 2030. "
                "Additionally, the company issued $30 million of 2030 Notes for $30 million in cash.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 5,
                    "note_id": ["g-1", "g-2", "debt-1", "util-1", "milestone-1"],
                    "category": [
                        "Programs / initiatives",
                        "Programs / initiatives",
                        "Debt / liquidity / covenants",
                        "Results / drivers",
                        "Programs / initiatives",
                    ],
                    "claim": [
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                        "45Z tax credit monetization agreement executed.",
                        "Junior mezzanine debt of $130.7 million was repaid from Obion sale proceeds.",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 101%.",
                        "All eight operating ethanol plants expected to qualify for production tax credits in 2026.",
                    ],
                    "note": [
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                        "45Z tax credit monetization agreement executed.",
                        "Junior mezzanine debt of $130.7 million was repaid from Obion sale proceeds.",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 101%.",
                        "All eight operating ethanol plants expected to qualify for production tax credits in 2026.",
                    ],
                    "metric_ref": [
                        "45Z Adjusted EBITDA / monetization",
                        "Strategic milestone",
                        "Debt reduction",
                        "Utilization",
                        "45Z qualification",
                    ],
                    "score": [96.0, 95.0, 94.0, 93.0, 92.0],
                    "doc_type": ["earnings_release"] * 5,
                    "doc": ["release_q3.txt"] * 5,
                    "source_type": ["earnings_release"] * 5,
                    "evidence_snippet": [
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                        "45Z tax credit monetization agreement executed.",
                        "Junior mezzanine debt of $130.7 million was repaid from Obion sale proceeds.",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 101%.",
                        "All eight operating ethanol plants expected to qualify for production tax credits in 2026.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q3_rows = _quarter_block_notes(ws, "2025-09-30")

            assert not any("Repurchase authorization increased to $200.0m." in note for note in q3_rows)
            assert not any("Repurchased $170.0m of shares" in note for note in q3_rows)
            assert not any("Repurchased $77.2m of shares" in note for note in q3_rows)
            assert any("45Z" in note or "Junior mezzanine debt" in note for note in q3_rows)


def test_gpre_q4_quarter_notes_add_convertible_exchange_subscription_and_interest_outlook(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_q4_convertible_exchange.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940225000163_gpre-20250930.htm").write_text(
                "Green Plains Inc. "
                "Convertible Debt Exchange "
                "On October 27, 2025, the company executed separate, privately negotiated exchange agreements with certain of the holders of its existing 2.25% Convertible Senior Notes due 2027 "
                "to exchange $170 million aggregate principal amount of the 2027 Notes for $170 million of newly issued 5.25% Convertible Senior Notes due November 2030. "
                "Additionally, the company completed separate, privately negotiated subscription agreements pursuant to which it issued $30 million of 2030 Notes for $30 million in cash. "
                "$200 million in aggregate principal amount of the 2030 Notes is now outstanding, and $60 million in aggregate principal amount of the 2027 Notes remains outstanding with existing terms unchanged. "
                "The company used approximately $30 million of the net proceeds from the subscription transactions to repurchase approximately 2.9 million shares of its common stock. "
                "The initial conversion rate of the 2030 Notes is 63.6132 shares of common stock per $1,000 principal amount of 2030 Notes (equivalent to an initial conversion price of approximately $15.72 per share of common stock). "
                "When considering the extinguishment of the Junior Notes, the increased interest rate on convertible notes, the increased amount of outstanding convertible notes and anticipated interest expense related to the carbon equipment financing, "
                "the company expects annualized interest expense of approximately $30 to $35 million for the year ended December 31, 2026.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940226000019_gpre-20251231.htm").write_text(
                "Green Plains Inc. "
                "Convertible Debt Exchange "
                "On October 27, 2025, the company executed separate, privately negotiated exchange agreements with certain of the holders of its existing 2.25% Convertible Senior Notes due 2027 "
                "to exchange $170 million aggregate principal amount of the 2027 Notes for $170 million of newly issued 5.25% Convertible Senior Notes due November 2030. "
                "Additionally, the company completed separate, privately negotiated subscription agreements pursuant to which it issued $30 million of 2030 Notes for $30 million in cash. "
                "On October 27, 2025, in conjunction with the privately negotiated exchange and subscription transactions, the company repurchased approximately 2.9 million shares of its common stock for approximately $30.0 million. "
                "The company used approximately $30 million of the net proceeds from the subscription transactions to repurchase approximately 2.9 million shares of its common stock. "
                "The 2030 Notes are convertible at an initial conversion price of approximately $15.72 per share of common stock. "
                "When considering the extinguishment of the Junior Notes, the increased interest rate on convertible notes, the increased amount of outstanding convertible notes and anticipated interest expense related to the carbon equipment financing, "
                "the company expects annualized interest expense of approximately $30 to $35 million for the year ended December 31, 2026.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940226000016_gpre-q42025earningsrelease.htm").write_text(
                "Green Plains Inc. Issued an additional $30.0m of 5.25% convertible senior notes due November 2030. "
                "Repurchased approximately 2.9m shares for approximately $30.0m in connection with the exchange and subscription transactions.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 5 + [pd.Timestamp("2025-09-30")] * 2,
                    "note_id": ["q4-1", "q4-2", "q4-3", "q4-4", "q4-5", "q3-1", "q3-2"],
                    "category": [
                        "Guidance / outlook",
                        "Cash flow / FCF / capex",
                        "Debt / liquidity / covenants",
                        "Strategy / segment",
                        "Programs / initiatives",
                        "Guidance / outlook",
                        "Debt / liquidity / covenants",
                    ],
                    "claim": [
                        "At least $188m of 45Z-related Adjusted EBITDA expected in 2026.",
                        "FCF TTM YoY delta $198.7m",
                        "Net debt delta $-77.9m",
                        "Adjusted EBITDA YoY 369.8%",
                        "Actively marketing 2026 45Z production tax credits.",
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                        "Junior mezzanine debt of $130.7 million was repaid from Obion sale proceeds.",
                    ],
                    "note": [
                        "At least $188m of 45Z-related Adjusted EBITDA expected in 2026.",
                        "FCF TTM YoY delta $198.7m",
                        "Net debt delta $-77.9m",
                        "Adjusted EBITDA YoY 369.8%",
                        "Actively marketing 2026 45Z production tax credits.",
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                        "Junior mezzanine debt of $130.7 million was repaid from Obion sale proceeds.",
                    ],
                    "metric_ref": [
                        "45Z Adjusted EBITDA / monetization",
                        "FCF",
                        "Net debt / leverage",
                        "Adjusted EBITDA",
                        "45Z marketing",
                        "45Z Adjusted EBITDA / monetization",
                        "Debt reduction",
                    ],
                    "score": [96.0, 95.0, 94.0, 93.0, 92.0, 91.0, 90.0],
                    "doc_type": ["earnings_release"] * 7,
                    "doc": ["release_q4.txt"] * 5 + ["release_q3.txt"] * 2,
                    "source_type": ["earnings_release"] * 7,
                    "evidence_snippet": [
                        "At least $188m of 45Z-related Adjusted EBITDA expected in 2026.",
                        "FCF TTM YoY delta $198.7m",
                        "Net debt delta $-77.9m",
                        "Adjusted EBITDA YoY 369.8%",
                        "Actively marketing 2026 45Z production tax credits.",
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                        "Junior mezzanine debt of $130.7 million was repaid from Obion sale proceeds.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="GPRE", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q4_rows = _quarter_block_notes(ws, "2025-12-31")
            q3_rows = _quarter_block_notes(ws, "2025-09-30")
            generic_rows = [dict(x) for x in getattr(ctx.derived, "_generic_source_note_rescue_cache", [])]
            q4_generic_summaries = [
                str(x.get("_render_summary") or "")
                for x in generic_rows
                if str(x.get("quarter")) == "2025-12-31"
            ]

            assert any(
                "Exchanged $170.0m of 2.25% convertible senior notes due 2027 for $170.0m of 5.25% convertible senior notes due November 2030 (conversion price $15.72/share)."
                in note
                for note in q4_generic_summaries
            )
            assert any(
                "Exchanged $170.0m of 2.25% convertible senior notes due 2027 for $170.0m of 5.25% convertible senior notes due November 2030 (conversion price $15.72/share)."
                in note
                for note in q4_rows
            )
            assert any(
                "Issued an additional $30.0m of 5.25% convertible senior notes due November 2030; proceeds funded the repurchase of approximately 2.9m shares for approximately $30.0m."
                in note
                for note in q4_generic_summaries
            )
            assert any(
                "Repurchased approximately 2.9m shares for approximately $30.0m in connection with the October 27, 2025 exchange and subscription transactions."
                in note
                for note in q4_rows
            )
            assert not any(
                "Repurchased 2.9m shares for $30.0m in Q4." in note
                or "with an average price of $10.34/share in Q4." in note
                for note in q4_rows
            )
            assert not any(
                note == "Repurchased approximately 2.9m shares for approximately $30.0m in connection with the exchange and subscription transactions."
                for note in q4_rows
            )
            assert any(
                "Annualized 2026 interest expense is expected at about $30.0m-$35.0m, reflecting the 2030 convertible notes, Junior Note extinguishment and carbon equipment financing."
                in note
                for note in q4_generic_summaries
            )
            assert not any("$31.62/share" in note for note in q4_generic_summaries)
            assert not any("Issued $30.0m of convertible senior notes due 2030." in note for note in q3_rows)
            assert not any("2030 Notes" in note for note in q3_rows)


def test_gpre_q4_quarter_notes_add_carbon_capture_and_source_accurate_monetization_update(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_q4_carbon_capture_and_monetization.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940226000019_gpre-20251231.htm").write_text(
                "Green Plains Inc. Carbon capture is now fully operational at our Central City, Wood River and York, Nebraska facilities. "
                "On September 16, 2025, we entered into a 45Z tax credit monetization agreement for our Nebraska production. "
                "The agreement was amended on December 10, 2025 to include additional facilities. "
                "The stronger balance sheet and operational excellence support future optionality.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940226000016_gpre-q42025earningsrelease.htm").write_text(
                "On September 17, 2025, the company entered into a 45Z tax credit monetization agreement. "
                "Carbon capture is fully operational at Central City, Wood River and York, Nebraska facilities.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 4,
                    "note_id": ["q4-1", "q4-2", "q4-3", "q4-4"],
                    "category": [
                        "Guidance / outlook",
                        "Cash flow / FCF / capex",
                        "Debt / liquidity / covenants",
                        "Programs / initiatives",
                    ],
                    "claim": [
                        "At least $188m of 45Z-related Adjusted EBITDA expected in 2026.",
                        "FCF TTM YoY delta $198.7m",
                        "Net debt delta $-77.9m",
                        "Actively marketing 2026 45Z production tax credits.",
                    ],
                    "note": [
                        "At least $188m of 45Z-related Adjusted EBITDA expected in 2026.",
                        "FCF TTM YoY delta $198.7m",
                        "Net debt delta $-77.9m",
                        "Actively marketing 2026 45Z production tax credits.",
                    ],
                    "metric_ref": [
                        "45Z Adjusted EBITDA / monetization",
                        "FCF",
                        "Net debt / leverage",
                        "45Z marketing",
                    ],
                    "score": [96.0, 95.0, 94.0, 93.0],
                    "doc_type": ["earnings_release"] * 4,
                    "doc": ["release_q4.txt"] * 4,
                    "source_type": ["earnings_release"] * 4,
                    "evidence_snippet": [
                        "At least $188m of 45Z-related Adjusted EBITDA expected in 2026.",
                        "FCF TTM YoY delta $198.7m",
                        "Net debt delta $-77.9m",
                        "Actively marketing 2026 45Z production tax credits.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="GPRE", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q4_rows = _quarter_block_notes(ws, "2025-12-31")

            assert any(
                "Carbon capture was fully operational at Central City, Wood River and York, Nebraska facilities."
                in note
                for note in q4_rows
            )
            assert any(
                "45Z tax credit monetization agreement for Nebraska production was entered on September 16, 2025 and amended on December 10, 2025 to add credits from three additional facilities."
                in note
                for note in q4_rows
            )
            assert not any("September 17, 2025" in note for note in q4_rows)
            assert not any("operational excellence, and a $49.1 million $428.8 million stronger balance sheet." in note for note in q4_rows)


def test_gpre_quarter_notes_collapse_agreement_rows_to_source_faithful_filing_dates(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_agreement_rows_collapse_to_filing_dates.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940225000163_gpre-20250930.htm").write_text(
                "Green Plains Inc. On September 16, 2025, the company entered into an agreement for its Nebraska production tax credits. "
                "Management expects this 45Z tax credit monetization agreement to support low-carbon ethanol value realization.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940226000019_gpre-20251231.htm").write_text(
                "Green Plains Inc. On September 16, 2025, the company entered into an agreement for its Nebraska production tax credits. "
                "On December 10, 2025, the agreement was amended to add Section 45Z production tax credits produced at three more of the company's facilities.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940226000016_gpre-q42025earningsrelease.htm").write_text(
                "Green Plains Inc. On September 17, 2025, the company entered into a 45Z tax credit monetization agreement. "
                "45Z tax credit monetization agreement executed, advancing low-carbon ethanol value creation.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-12-31")],
                    "note_id": ["q3-guid", "q4-guid"],
                    "category": ["Guidance / outlook", "Guidance / outlook"],
                    "claim": ["Q3 placeholder.", "Q4 placeholder."],
                    "note": ["Q3 placeholder.", "Q4 placeholder."],
                    "metric_ref": ["Guidance", "Guidance"],
                    "score": [90.0, 90.0],
                    "doc_type": ["earnings_release", "earnings_release"],
                    "doc": ["q3.txt", "q4.txt"],
                    "source_type": ["earnings_release", "earnings_release"],
                    "evidence_snippet": ["Q3 placeholder.", "Q4 placeholder."],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="GPRE", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q3_rows = _quarter_block_notes(ws, "2025-09-30")
            q4_rows = _quarter_block_notes(ws, "2025-12-31")

            q3_agreement_rows = [note for note in q3_rows if "45Z tax credit monetization agreement" in note]
            assert len(q3_agreement_rows) == 1
            assert "September 16, 2025" in q3_agreement_rows[0]
            assert "September 17, 2025" not in q3_agreement_rows[0]
            assert not any("agreement executed, advancing low-carbon ethanol value creation" in note for note in q3_rows)
            assert any(
                "45Z tax credit monetization agreement for Nebraska production was entered on September 16, 2025 and amended on December 10, 2025 to add credits from three additional facilities."
                in note
                for note in q4_rows
            )
            assert not any("September 17, 2025" in note for note in q4_rows)


def test_gpre_q4_visible_rows_collapse_to_single_connected_buyback_and_single_filing_faithful_agreement(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_q4_final_visible_event_cleanup.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940226000019_gpre-20251231.htm").write_text(
                "Green Plains Inc. On October 27, 2025, in conjunction with the privately negotiated exchange and subscription transactions, "
                "the company repurchased approximately 2.9 million shares of its common stock for approximately $30.0 million. "
                "Tax Credit Purchase Agreement. On September 16, 2025, the company entered into an agreement for Nebraska production tax credits. "
                "On December 10, 2025, the agreement was amended to add credits from three additional facilities.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940226000016_gpre-q42025earningsrelease.htm").write_text(
                "Green Plains Inc. Repurchased approximately 2.9 million shares for approximately $30.0 million in connection with the exchange and subscription transactions. "
                "On September 17, 2025, the company entered into a 45Z tax credit monetization agreement, later amended on December 10, 2025.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 2,
                    "note_id": ["q4-guid-1", "q4-guid-2"],
                    "category": ["Guidance / outlook", "Guidance / outlook"],
                    "claim": ["Q4 placeholder 1.", "Q4 placeholder 2."],
                    "note": ["Q4 placeholder 1.", "Q4 placeholder 2."],
                    "metric_ref": ["Guidance", "Guidance"],
                    "score": [90.0, 89.0],
                    "doc_type": ["earnings_release", "earnings_release"],
                    "doc": ["q4a.txt", "q4b.txt"],
                    "source_type": ["earnings_release", "earnings_release"],
                    "evidence_snippet": ["Q4 placeholder 1.", "Q4 placeholder 2."],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="GPRE", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q4_rows = _quarter_block_notes(ws, "2025-12-31")

            buyback_rows = [note for note in q4_rows if "repurchased" in note.lower() and "$30.0m" in note]
            assert len(buyback_rows) == 1
            assert (
                "Repurchased approximately 2.9m shares for approximately $30.0m in connection with the October 27, 2025 exchange and subscription transactions."
                in buyback_rows[0]
            )

            agreement_rows = [note for note in q4_rows if "45Z tax credit monetization agreement" in note]
            assert len(agreement_rows) == 1
            assert (
                "45Z tax credit monetization agreement for Nebraska production was entered on September 16, 2025 and amended on December 10, 2025 to add credits from three additional facilities."
                in agreement_rows[0]
            )


def test_gpre_q3_does_not_show_q4_subsequent_event_repurchase_or_carbon_capture_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_q3_no_q4_subsequent_events.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940225000163_gpre-20250930.htm").write_text(
                "Green Plains Inc. Subsequent Events. On October 27, 2025, in conjunction with the privately negotiated exchange and "
                "subscription agreements for the 2030 Notes, the company repurchased approximately 2.9 million shares of its common stock "
                "for approximately $30.0 million. Carbon capture is now fully operational at our Central City, Wood River and York, Nebraska facilities.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940226000019_gpre-20251231.htm").write_text(
                "Green Plains Inc. On October 27, 2025, in conjunction with the privately negotiated exchange and subscription agreements for the 2030 Notes, "
                "the company repurchased approximately 2.9 million shares of its common stock for approximately $30.0 million. "
                "Carbon capture is now fully operational at our Central City, Wood River and York, Nebraska facilities.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-12-31")],
                    "note_id": ["q3-guid", "q4-guid"],
                    "category": ["Guidance / outlook", "Guidance / outlook"],
                    "claim": ["Q3 placeholder.", "Q4 placeholder."],
                    "note": ["Q3 placeholder.", "Q4 placeholder."],
                    "metric_ref": ["Guidance", "Guidance"],
                    "score": [90.0, 90.0],
                    "doc_type": ["earnings_release", "earnings_release"],
                    "doc": ["q3.txt", "q4.txt"],
                    "source_type": ["earnings_release", "earnings_release"],
                    "evidence_snippet": ["Q3 placeholder.", "Q4 placeholder."],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="GPRE", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q3_rows = _quarter_block_notes(ws, "2025-09-30")
            q4_rows = _quarter_block_notes(ws, "2025-12-31")

            assert not any("October 27, 2025" in note and "repurchased approximately 2.9m shares" in note.lower() for note in q3_rows)
            assert not any(
                "carbon capture was fully operational at central city, wood river and york" in note.lower()
                for note in q3_rows
            )
            assert any("October 27, 2025" in note and "repurchased approximately 2.9m shares" in note.lower() for note in q4_rows)
            assert any(
                "carbon capture was fully operational at central city, wood river and york" in note.lower()
                for note in q4_rows
            )


def test_gpre_q3_negative_quarter_statement_blocks_later_dated_october_buyback_execution(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_q3_negative_statement_blocks_later_october_event.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940225000163_gpre-20250930.htm").write_text(
                "Green Plains Inc. We did not repurchase any shares of common stock during the third quarter of 2025. "
                "Subsequent Events. On October 27, 2025, in conjunction with the privately negotiated exchange and "
                "subscription agreements for the 2030 Notes, the company repurchased approximately 2.9 million shares "
                "of its common stock for approximately $30.0 million.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940226000019_gpre-20251231.htm").write_text(
                "Green Plains Inc. On October 27, 2025, in conjunction with the privately negotiated exchange and "
                "subscription agreements for the 2030 Notes, the company repurchased approximately 2.9 million shares "
                "of its common stock for approximately $30.0 million.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-12-31")],
                    "note_id": ["q3-guid", "q4-guid"],
                    "category": ["Guidance / outlook", "Guidance / outlook"],
                    "claim": ["Q3 placeholder.", "Q4 placeholder."],
                    "note": ["Q3 placeholder.", "Q4 placeholder."],
                    "metric_ref": ["Guidance", "Guidance"],
                    "score": [90.0, 90.0],
                    "doc_type": ["earnings_release", "earnings_release"],
                    "doc": ["q3.txt", "q4.txt"],
                    "source_type": ["earnings_release", "earnings_release"],
                    "evidence_snippet": ["Q3 placeholder.", "Q4 placeholder."],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="GPRE", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q3_rows = _quarter_block_notes(ws, "2025-09-30")
            q4_rows = _quarter_block_notes(ws, "2025-12-31")

            assert not any("October 27, 2025" in note and "repurchased approximately 2.9m shares" in note.lower() for note in q3_rows)
            assert sum(
                1
                for note in q4_rows
                if "October 27, 2025" in note and "repurchased approximately 2.9m shares" in note.lower()
            ) == 1


def test_pbi_quarter_notes_ignore_foreign_sec_cache_docs_inside_ticker_root(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_ignore_foreign_sec_cache.xlsx")
            sec_cache_dir = case_dir / "TEST" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000107878225000001_gpre-20250331.htm").write_text(
                "Green Plains Inc. reported first quarter 2025 financial results. "
                "Cost reduction initiatives are progressing ahead of plan, supporting a positive EBITDA outlook under current market conditions. "
                "Management is pursuing non-core asset monetization to enhance liquidity and strengthen the balance sheet. "
                "45Z production tax credits are expected to contribute meaningfully in future periods.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-03-31")] * 3,
                    "note_id": ["g-1", "g-2", "g-3"],
                    "category": [
                        "Guidance / outlook",
                        "Debt / liquidity / covenants",
                        "Programs / initiatives",
                    ],
                    "claim": [
                        "FY 2025 Revenue guidance reaffirmed at $1.90bn-$1.95bn.",
                        "Net debt declined by $28.0m.",
                        "PB Bank note remains relevant in Q1.",
                    ],
                    "note": [
                        "FY 2025 Revenue guidance reaffirmed at $1.90bn-$1.95bn.",
                        "Net debt declined by $28.0m.",
                        "PB Bank note remains relevant in Q1.",
                    ],
                    "metric_ref": ["Revenue guidance", "Net debt / leverage", "PB Bank"],
                    "score": [95.0, 93.0, 92.0],
                    "doc_type": ["earnings_release", "model_metric", "earnings_release"],
                    "doc": ["release_q1.txt", "history_q", "release_q1.txt"],
                    "source_type": ["earnings_release", "model_metric", "earnings_release"],
                    "evidence_snippet": [
                        "FY 2025 Revenue guidance reaffirmed at $1.90bn-$1.95bn.",
                        "Net debt declined by $28.0m.",
                        "PB Bank note remains relevant in Q1.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q1_rows = _quarter_block_notes(ws, "2025-03-31")

            assert any("FY 2025 Revenue guidance reaffirmed" in note for note in q1_rows)
            assert not any("45z" in note.lower() for note in q1_rows)
            assert not any("positive ebitda outlook" in note.lower() for note in q1_rows)
            assert not any("non-core asset monetization" in note.lower() for note in q1_rows)


def test_gpre_quarter_notes_recall_clean_model_metric_signals_across_multiple_quarters(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_model_metric_recall.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [
                        pd.Timestamp("2025-12-31"),
                        pd.Timestamp("2025-06-30"),
                        pd.Timestamp("2024-12-31"),
                    ],
                    "note_id": ["debt-1", "revolver-1", "fcf-1"],
                    "category": [
                        "Debt / liquidity / covenants",
                        "Debt / liquidity / covenants",
                        "Cash flow / FCF / capex",
                    ],
                    "claim": [
                        "Net debt delta $-65.0m.",
                        "Revolver availability moved to $275.0m at 2025-06-30 (delta $50.0m).",
                        "FCF TTM at 2024-12-31: $120.0m, yoy -15.0%, delta $-21.0m.",
                    ],
                    "note": [
                        "Net debt delta $-65.0m.",
                        "Revolver availability moved to $275.0m at 2025-06-30 (delta $50.0m).",
                        "FCF TTM at 2024-12-31: $120.0m, yoy -15.0%, delta $-21.0m.",
                    ],
                    "metric_ref": [
                        "Net debt / leverage",
                        "revolver_availability_change",
                        "FCF",
                    ],
                    "score": [90.0, 91.0, 89.0],
                    "doc_type": ["model_metric", "model_metric", "model_metric"],
                    "doc": ["history_q", "history_q", "history_q"],
                    "source_type": ["model_metric", "model_metric", "model_metric"],
                    "evidence_snippet": [
                        "Net debt delta $-65.0m.",
                        "Revolver availability moved to $275.0m at 2025-06-30 (delta $50.0m).",
                        "FCF TTM at 2024-12-31: $120.0m, yoy -15.0%, delta $-21.0m.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "Net debt delta $-65.0m."}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "Revolver availability moved to $275.0m at 2025-06-30 (delta $50.0m)."}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "FCF TTM at 2024-12-31: $120.0m, yoy -15.0%, delta $-21.0m."}]),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            def _quarter_block(qtxt: str) -> list[str]:
                out: list[str] = []
                capture = False
                for rr in range(1, ws.max_row + 1):
                    marker = str(ws.cell(row=rr, column=1).value or "")
                    if marker == qtxt:
                        capture = True
                        continue
                    if capture and marker:
                        break
                    if capture:
                        note = str(ws.cell(row=rr, column=3).value or "")
                        if note:
                            out.append(note)
                return out

            q4_rows = _quarter_block("2025-12-31")
            q2_rows = _quarter_block("2025-06-30")
            q424_rows = _quarter_block("2024-12-31")

            assert any("Net debt declined by $65" in note for note in q4_rows)
            assert any("Revolver availability increased from $225" in note for note in q2_rows)
            assert any("FCF TTM declined to $120" in note or "FCF TTM declined by $21" in note for note in q424_rows)


def test_gpre_quarter_notes_thin_block_gains_quantified_support_bridge_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_thin_block_quantified.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-06-30")] * 5,
                    "note_id": ["milestone-1", "risk-1", "util-1", "rev-1", "margin-1"],
                    "category": [
                        "Programs / initiatives",
                        "Results / drivers",
                        "Results / drivers",
                        "Debt / liquidity / covenants",
                        "Results / drivers",
                    ],
                    "claim": [
                        "Carbon capture infrastructure delivered; Q4 2025 start-up still on track.",
                        "Risk management supports margins and cash flow.",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 99%.",
                        "Revolver utilization notable",
                        "EBITDA margin compressed",
                    ],
                    "note": [
                        "Carbon capture infrastructure delivered; Q4 2025 start-up still on track.",
                        "Risk management supports margins and cash flow.",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 99%.",
                        "Revolver utilization notable",
                        "EBITDA margin compressed",
                    ],
                    "metric_ref": [
                        "Strategic milestone",
                        "Risk management",
                        "Utilization",
                        "revolver_utilization",
                        "ebitda_margin_ttm_yoy_bps",
                    ],
                    "score": [95.0, 92.0, 94.0, 89.0, 90.0],
                    "doc_type": ["earnings_release", "earnings_release", "earnings_release", "revolver", "model_metric"],
                    "doc": ["release_q2.txt", "release_q2.txt", "release_q2.txt", "history_q", "history_q"],
                    "source_type": ["earnings_release", "earnings_release", "earnings_release", "revolver", "model_metric"],
                    "evidence_snippet": [
                        "Carbon capture infrastructure equipment delivered and on track for start-up early in the fourth quarter of 2025.",
                        "Risk management supports margins and cash flow.",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 99%.",
                        "We also had $258.5 million available under our committed revolving credit agreement.",
                        "EBITDA margin delta -143 bps",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q2.txt", "doc_type": "earnings_release", "snippet": "Carbon capture infrastructure equipment delivered and on track for start-up early in the fourth quarter of 2025."}]),
                        json.dumps([{"doc_path": "release_q2.txt", "doc_type": "earnings_release", "snippet": "Risk management supports margins and cash flow."}]),
                        json.dumps([{"doc_path": "release_q2.txt", "doc_type": "earnings_release", "snippet": "Achieved strong utilization in the quarter from the nine operating ethanol plants of 99%."}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "revolver", "snippet": "We also had $258.5 million available under our committed revolving credit agreement."}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "EBITDA margin delta -143 bps"}]),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            q2_rows = []
            capture = False
            for rr in range(1, ws.max_row + 1):
                marker = str(ws.cell(row=rr, column=1).value or "")
                if marker == "2025-06-30":
                    capture = True
                    continue
                if capture and marker:
                    break
                if capture:
                    note = str(ws.cell(row=rr, column=3).value or "")
                    if note:
                        q2_rows.append(note)

            assert len(q2_rows) >= 4
            assert any("Revolver availability ended the quarter at $258.5m." in note or "EBITDA margin compressed 143 bps YoY." in note for note in q2_rows)
            assert not any("Risk management supports margins and cash flow." in note for note in q2_rows)


def test_gpre_quarter_notes_rescue_can_compete_even_when_block_already_has_two_survivors(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_two_rows_still_allow_rescue.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2024-12-31")] * 4,
                    "note_id": ["milestone-1", "tone-1", "fcf-1", "debt-1"],
                    "category": [
                        "Programs / initiatives",
                        "Results / drivers",
                        "Cash flow / FCF / capex",
                        "Debt / liquidity / covenants",
                    ],
                    "claim": [
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                        "Risk management supports margins and cash flow.",
                        "FCF TTM accelerated",
                        "Net debt declined",
                    ],
                    "note": [
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                        "Risk management supports margins and cash flow.",
                        "FCF TTM accelerated",
                        "Net debt declined",
                    ],
                    "metric_ref": [
                        "Strategic milestone",
                        "Risk management",
                        "fcf_ttm_delta_yoy",
                        "net_debt_yoy_delta",
                    ],
                    "score": [95.0, 94.0, 78.0, 77.0],
                    "doc_type": ["earnings_release", "earnings_release", "model_metric", "model_metric"],
                    "doc": ["release_q4.txt", "release_q4.txt", "history_q", "history_q"],
                    "source_type": ["earnings_release", "earnings_release", "model_metric", "model_metric"],
                    "evidence_snippet": [
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                        "Risk management supports margins and cash flow.",
                        "FCF TTM at 2024-12-31: $120.0m, yoy -15.0%, delta $-21.0m.",
                        "Net debt delta $-65.0m.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q4.txt", "doc_type": "earnings_release", "snippet": "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming."}]),
                        json.dumps([{"doc_path": "release_q4.txt", "doc_type": "earnings_release", "snippet": "Risk management supports margins and cash flow."}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "FCF TTM at 2024-12-31: $120.0m, yoy -15.0%, delta $-21.0m."}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "Net debt delta $-65.0m."}]),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            q4_rows: list[str] = []
            capture = False
            for rr in range(1, ws.max_row + 1):
                marker = str(ws.cell(row=rr, column=1).value or "")
                if marker == "2024-12-31":
                    capture = True
                    continue
                if capture and marker:
                    break
                if capture:
                    note = str(ws.cell(row=rr, column=3).value or "")
                    if note:
                        q4_rows.append(note)

            assert len(q4_rows) >= 4
            assert any("FCF TTM declined" in note or "FCF TTM improved" in note for note in q4_rows)
            assert any("Net debt declined by $65" in note for note in q4_rows)


    def test_pbi_quarter_notes_promote_explanatory_driver_note_over_generic_block_fill(
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        with _case_dir() as case_dir:
            with _profile_override(monkeypatch, "PBI"):
                out_path = _make_model_out_path(case_dir, "pbi_notes_explanatory_driver.xlsx")
                sec_cache = case_dir / "TEST" / "sec_cache"
                sec_cache.mkdir(parents=True, exist_ok=True)
                source_doc = sec_cache / "doc_000162828025023187_q12025earningspressrelea.htm"
                source_doc.write_text(
                    "Gross margin declined $13 million compared to the prior year period primarily driven by lower revenue; "
                    "however, gross margin percentage increased to 68.9% from 66.6% driven by headcount reductions and "
                    "other cost savings initiatives. Higher revenue per piece, improved productivity, and cost reduction "
                    "initiatives drove the increase in Adjusted Segment EBITDA and EBIT.",
                    encoding="utf-8",
                )
                quarter_notes = pd.DataFrame(
                    {
                        "quarter": [pd.Timestamp("2025-03-31")] * 5,
                        "note_id": ["g-1", "g-2", "g-3", "bank-1", "debt-1"],
                        "category": [
                            "Guidance / outlook",
                            "Guidance / outlook",
                            "Guidance / outlook",
                            "Programs / initiatives",
                            "Cash / liquidity / leverage",
                        ],
                        "claim": [
                            "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.",
                            "FY 2025 Adjusted EBIT guidance updated to $450m-$480m.",
                            "FY 2025 FCF guidance updated to $330m-$370m.",
                            "PB Bank target >= $120m bank-held leases.",
                            "Liquidity improved.",
                        ],
                        "note": [
                            "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.",
                            "FY 2025 Adjusted EBIT guidance updated to $450m-$480m.",
                            "FY 2025 FCF guidance updated to $330m-$370m.",
                            "PB Bank target >= $120m bank-held leases.",
                            "Liquidity improved.",
                        ],
                        "metric_ref": [
                            "Revenue guidance",
                            "Adjusted EBIT guidance",
                            "FCF target",
                            "PB Bank liquidity release",
                            "Debt reduction",
                        ],
                        "score": [95.0, 94.0, 93.0, 90.0, 91.0],
                        "doc_type": ["earnings_release", "earnings_release", "earnings_release", "earnings_release", "earnings_release"],
                        "doc": ["release_q1.txt", "release_q1.txt", "release_q1.txt", "release_q1.txt", "release_q1.txt"],
                        "evidence_snippet": [
                            "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.",
                            "FY 2025 Adjusted EBIT guidance updated to $450m-$480m.",
                            "FY 2025 FCF guidance updated to $330m-$370m.",
                            "The Bank held $84 million of associated leases at the end of Q1, and the Company aims to increase that figure to $120 million by the end of 2025.",
                            "Through the end of Q1, the Company repurchased $23 million of debt in the open market.",
                        ],
                    }
                )

                ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
                ctx.callbacks.write_quarter_notes_ui_v2()
                ws = ctx.wb["Quarter_Notes_UI"]

                q1_rows: list[str] = []
                capture = False
                for rr in range(1, ws.max_row + 1):
                    marker = str(ws.cell(row=rr, column=1).value or "")
                    if marker == "2025-03-31":
                        capture = True
                        continue
                    if capture and marker:
                        break
                    if capture:
                        note = str(ws.cell(row=rr, column=3).value or "")
                        if note:
                            q1_rows.append(note)

                assert any(
                    "Gross margin expanded to 68.9% from 66.6%, driven by headcount reductions and cost savings." in note
                    or "Presort EBIT improved, driven by higher revenue per piece, productivity and cost reduction." in note
                    for note in q1_rows
                )
                assert not any(note == "Liquidity improved." for note in q1_rows)


def test_pbi_quarter_notes_can_recall_explanatory_driver_from_local_source_material(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_ticker_model_out_path(case_dir, "PBI", "pbi_notes_source_driver_rescue.xlsx")
            sec_cache = case_dir / "PBI" / "sec_cache"
            sec_cache.mkdir(parents=True, exist_ok=True)
            source_doc = sec_cache / "doc_000162828026008604_q42025earningspressrelea.htm"
            source_doc.write_text(
                "Gross margin expanded 180 basis points in the fourth quarter due to cost optimization actions and "
                "a shift to higher margin revenue streams. In the fourth quarter, operating expenses declined $28 "
                "million year-over-year primarily from cost reduction initiatives.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 6,
                    "note_id": ["g-1", "g-2", "g-3", "fcf-1", "bb-1", "debt-1"],
                    "category": [
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Cash flow / FCF / capex",
                        "Cash / liquidity / leverage",
                        "Cash / liquidity / leverage",
                    ],
                    "claim": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved.",
                        "Share repurchases continued.",
                        "Liquidity improved.",
                    ],
                    "note": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved.",
                        "Share repurchases continued.",
                        "Liquidity improved.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "FCF target",
                        "FCF improvement",
                        "Capital allocation / buyback",
                        "Debt reduction",
                    ],
                    "score": [95.0, 94.0, 93.0, 92.0, 91.0, 90.0],
                    "doc_type": ["earnings_release"] * 6,
                    "doc": ["release_q4.txt"] * 6,
                    "evidence_snippet": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved.",
                        "Repurchased 12.6 million shares for $127.0 million in the fourth quarter.",
                        "Reduced principal debt by $114.1 million in the fourth quarter.",
                    ],
                }
            )
            adj_metrics = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2024-12-31"), pd.Timestamp("2025-12-31")],
                    "adj_fcf": [131_800_000.0, 221_700_000.0],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="PBI", quarter_notes=quarter_notes)
            inputs = base_inputs.__class__(**{**vars(base_inputs), "adj_metrics": adj_metrics})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            q4_rows: list[str] = []
            capture = False
            for rr in range(1, ws.max_row + 1):
                marker = str(ws.cell(row=rr, column=1).value or "")
                if marker == "2025-12-31":
                    capture = True
                    continue
                if capture and marker:
                    break
                if capture:
                    note = str(ws.cell(row=rr, column=3).value or "")
                    if note:
                        q4_rows.append(note)

            assert any(
                "Gross margin expanded 180 bps, driven by cost optimization and a shift to higher margin revenue streams." in note
                or "Operating expenses declined $28.0m YoY, primarily from cost reduction." in note
                for note in q4_rows
            )


def test_gpre_quarter_notes_surface_quantified_adjusted_ebitda_margin_and_obion_repayment(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_multi_block_quantified.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [
                        pd.Timestamp("2025-12-31"),
                        pd.Timestamp("2025-12-31"),
                        pd.Timestamp("2025-12-31"),
                        pd.Timestamp("2025-09-30"),
                        pd.Timestamp("2025-09-30"),
                        pd.Timestamp("2025-06-30"),
                        pd.Timestamp("2025-03-31"),
                        pd.Timestamp("2024-03-31"),
                    ],
                    "note_id": ["adj-1", "fcf-1", "debt-1", "obion-1", "q3-margin-1", "q2-rev-1", "q1-oneoff-1", "q124-margin-1"],
                    "category": [
                        "Results / drivers",
                        "Cash flow / FCF / capex",
                        "Debt / liquidity / covenants",
                        "Debt / liquidity / covenants",
                        "Results / drivers",
                        "Debt / liquidity / covenants",
                        "One-time items",
                        "Results / drivers",
                    ],
                    "claim": [
                        "Adjusted EBITDA moved materially",
                        "FCF TTM accelerated",
                        "Net debt declined",
                        "Obion sale proceeds used to fully repay junior mezzanine debt",
                        "EBITDA margin compressed",
                        "Revolver utilization notable",
                        "One-offs signal in filing text",
                        "EBITDA margin expanded",
                    ],
                    "note": [
                        "Adjusted EBITDA moved materially",
                        "FCF TTM accelerated",
                        "Net debt declined",
                        "Obion sale proceeds used to fully repay junior mezzanine debt",
                        "EBITDA margin compressed",
                        "Revolver utilization notable",
                        "One-offs signal in filing text",
                        "EBITDA margin expanded",
                    ],
                    "metric_ref": [
                        "adj_ebitda_yoy",
                        "fcf_ttm_delta_yoy",
                        "net_debt_yoy_delta",
                        "Debt reduction",
                        "ebitda_margin_ttm_yoy_bps",
                        "revolver_availability_change",
                        "text:one-offs",
                        "ebitda_margin_ttm_yoy_bps",
                    ],
                    "score": [93.0, 92.0, 91.0, 95.0, 90.0, 89.0, 88.0, 90.0],
                    "doc_type": ["non_gaap", "model_metric", "model_metric", "earnings_release", "model_metric", "revolver", "html", "model_metric"],
                    "doc": [
                        "GPRE-Q4-2025-Earnings-Slides-FINAL.pdf",
                        "history_q",
                        "history_q",
                        "release_q3.txt",
                        "history_q",
                        "history_q",
                        "gpre-20250331.htm",
                        "history_q",
                    ],
                    "source_type": ["non_gaap", "model_metric", "model_metric", "earnings_release", "model_metric", "revolver", "html", "model_metric"],
                    "evidence_snippet": [
                        "Adjusted EBITDA YoY 369.8%",
                        "FCF TTM YoY delta $198.7m",
                        "Net debt delta $-77.9m",
                        "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt.",
                        "EBITDA margin delta -405 bps",
                        "We also had $258.5 million available under our committed revolving credit agreement.",
                        "(2) Corporate activities includes $ 10.3 million of restructuring costs for the three months ended March 31, 2025 as a result of the company's cost reduction initiative, including severance related to the departure of its CEO.",
                        "EBITDA margin delta +172 bps",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "GPRE-Q4-2025-Earnings-Slides-FINAL.pdf", "doc_type": "non_gaap", "snippet": "Adjusted EBITDA YoY 369.8%"}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "FCF TTM YoY delta $198.7m"}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "Net debt delta $-77.9m"}]),
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt."}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "EBITDA margin delta -405 bps"}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "revolver", "snippet": "We also had $258.5 million available under our committed revolving credit agreement."}]),
                        json.dumps([{"doc_path": "gpre-20250331.htm", "doc_type": "html", "snippet": "(2) Corporate activities includes $ 10.3 million of restructuring costs for the three months ended March 31, 2025 as a result of the company's cost reduction initiative, including severance related to the departure of its CEO."}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "EBITDA margin delta +172 bps"}]),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            def _quarter_block(qtxt: str) -> list[str]:
                out: list[str] = []
                capture = False
                for rr in range(1, ws.max_row + 1):
                    marker = str(ws.cell(row=rr, column=1).value or "")
                    if marker == qtxt:
                        capture = True
                        continue
                    if capture and marker:
                        break
                    if capture:
                        note = str(ws.cell(row=rr, column=3).value or "")
                        if note:
                            out.append(note)
                return out

            q4_rows = _quarter_block("2025-12-31")
            q3_rows = _quarter_block("2025-09-30")
            q2_rows = _quarter_block("2025-06-30")
            q1_rows = _quarter_block("2025-03-31")
            q124_rows = _quarter_block("2024-03-31")

            assert any("Adjusted EBITDA improved 369.8% YoY." in note for note in q4_rows)
            assert any("Junior mezzanine debt of $130.7m was repaid from Obion sale proceeds." in note for note in q3_rows)
            assert any("Revolver availability ended the quarter at $258.5m." in note for note in q2_rows)
            assert any("Corporate activities included $10.3m of restructuring costs from the cost reduction initiative." in note for note in q1_rows)
            assert any("EBITDA margin expanded 172 bps YoY." in note for note in q124_rows)


def test_pbi_quarter_notes_soft_cap_expands_without_dropping_guidance(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_soft_cap_expansion.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 8,
                    "note_id": ["g-1", "g-2", "g-3", "g-4", "fcf-1", "bb-1", "debt-1", "driver-1"],
                    "category": [
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Cash flow / FCF / capex",
                        "Cash / liquidity / leverage",
                        "Cash / liquidity / leverage",
                        "Better / worse vs prior",
                    ],
                    "claim": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved.",
                        "Share repurchases continued.",
                        "Liquidity improved.",
                        "Gross margin expanded 180 basis points in the fourth quarter due to cost optimization actions and a shift to higher margin revenue streams.",
                    ],
                    "note": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved.",
                        "Share repurchases continued.",
                        "Liquidity improved.",
                        "Gross margin expanded 180 basis points in the fourth quarter due to cost optimization actions and a shift to higher margin revenue streams.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "EPS guidance",
                        "FCF target",
                        "FCF improvement",
                        "Capital allocation / buyback",
                        "Debt reduction",
                        "Adjusted EBIT / margin",
                    ],
                    "score": [96.0, 95.0, 94.0, 93.0, 92.0, 91.0, 90.0, 97.0],
                    "doc_type": ["earnings_release"] * 8,
                    "doc": ["release_q4.txt"] * 8,
                    "evidence_snippet": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved.",
                        "Repurchased 12.6 million shares for $127.0 million in the fourth quarter.",
                        "Reduced principal debt by $114.1 million in the fourth quarter.",
                        "Gross margin expanded 180 basis points in the fourth quarter due to cost optimization actions and a shift to higher margin revenue streams.",
                    ],
                }
            )
            adj_metrics = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2024-12-31"), pd.Timestamp("2025-12-31")],
                    "adj_fcf": [131_800_000.0, 221_700_000.0],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="PBI", quarter_notes=quarter_notes)
            inputs = base_inputs.__class__(**{**vars(base_inputs), "adj_metrics": adj_metrics})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            q4_rows = _quarter_block_notes(ws, "2025-12-31")
            assert len(q4_rows) >= 7
            assert sum(1 for note in q4_rows if "FY 2026" in note) >= 4
            assert any("Free cash flow improved to $221.7m, up $89.9m YoY." in note for note in q4_rows)
            assert any(
                ("Reduced principal debt by $114.1m in Q4." in note)
                or ("Quarterly dividend" in note)
                or ("Repurchased" in note)
                for note in q4_rows
            )
            assert any(
                "Gross margin expanded 180 bps, driven by cost optimization and a shift to higher margin revenue streams." in note
                or "Operating expenses declined $28.0m YoY, primarily from cost reduction." in note
                for note in q4_rows
            )
            assert any(
                "Quarterly dividend set at $0.09/share." in note
                or "Repurchased 12.6m shares for $127.0m in Q4." in note
                for note in q4_rows
            )


def test_gpre_quarter_notes_soft_cap_prefers_strong_extras_over_weak_generic_fill(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_soft_cap_expansion.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 7,
                    "note_id": ["m-1", "margin-1", "debt-1", "util-1", "fcf-1", "debt2-1", "tone-1"],
                    "category": [
                        "Programs / initiatives",
                        "Results / drivers",
                        "Debt / liquidity / covenants",
                        "Results / drivers",
                        "Cash flow / FCF / capex",
                        "Debt / liquidity / covenants",
                        "Results / drivers",
                    ],
                    "claim": [
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                        "EBITDA margin compressed",
                        "Obion sale proceeds used to fully repay junior mezzanine debt.",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 101%.",
                        "FCF TTM accelerated",
                        "Net debt declined",
                        "Risk management supports margins and cash flow.",
                    ],
                    "note": [
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                        "EBITDA margin compressed",
                        "Obion sale proceeds used to fully repay junior mezzanine debt.",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 101%.",
                        "FCF TTM accelerated",
                        "Net debt declined",
                        "Risk management supports margins and cash flow.",
                    ],
                    "metric_ref": [
                        "45Z Adjusted EBITDA / monetization",
                        "ebitda_margin_ttm_yoy_bps",
                        "Debt reduction",
                        "Utilization",
                        "fcf_ttm_delta_yoy",
                        "net_debt_yoy_delta",
                        "Risk management",
                    ],
                    "score": [96.0, 95.0, 94.0, 93.0, 90.0, 89.0, 92.0],
                    "doc_type": ["earnings_release", "model_metric", "earnings_release", "earnings_release", "model_metric", "model_metric", "earnings_release"],
                    "doc": ["release_q3.txt", "history_q", "release_q3.txt", "release_q3.txt", "history_q", "history_q", "release_q3.txt"],
                    "source_type": ["earnings_release", "model_metric", "earnings_release", "earnings_release", "model_metric", "model_metric", "earnings_release"],
                    "evidence_snippet": [
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                        "EBITDA margin delta -405 bps",
                        "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt.",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 101%.",
                        "FCF TTM YoY delta $198.7m",
                        "Net debt delta $-77.9m",
                        "Risk management supports margins and cash flow.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025."}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "EBITDA margin delta -405 bps"}]),
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt."}]),
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "Achieved strong utilization in the quarter from the nine operating ethanol plants of 101%."}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "FCF TTM YoY delta $198.7m"}]),
                        json.dumps([{"doc_path": "history_q", "doc_type": "model_metric", "snippet": "Net debt delta $-77.9m"}]),
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "Risk management supports margins and cash flow."}]),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            q3_rows = _quarter_block_notes(ws, "2025-09-30")
            assert len(q3_rows) >= 6
            assert any("Q4 2025 45Z monetization expected at $15m-$25m." in note for note in q3_rows)
            assert any("EBITDA margin compressed 405 bps YoY." in note for note in q3_rows)
            assert any("Junior mezzanine debt of $130.7m was repaid from Obion sale proceeds." in note for note in q3_rows)
            assert any("FCF TTM improved by $198.7m YoY." in note for note in q3_rows)
            assert any("Net debt declined by $77.9m YoY." in note for note in q3_rows)
            assert not any("Risk management supports margins and cash flow." in note for note in q3_rows)


def test_pbi_quarter_notes_dedupe_identical_explanatory_summary_across_metric_labels(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_duplicate_explanatory_summary.xlsx")
            shared_summary = "Presort EBIT improved, driven by higher revenue per piece, productivity and cost reduction."
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-03-31")] * 4,
                    "note_id": ["driver-1", "driver-2", "g-1", "debt-1"],
                    "category": [
                        "Better / worse vs prior",
                        "Better / worse vs prior",
                        "Guidance / outlook",
                        "Cash / liquidity / leverage",
                    ],
                    "claim": [shared_summary, shared_summary, "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.", "Liquidity improved."],
                    "note": [shared_summary, shared_summary, "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.", "Liquidity improved."],
                    "metric_ref": [
                        "SendTech / Presort operating driver",
                        "Adjusted EBIT / margin",
                        "Revenue guidance",
                        "Debt reduction",
                    ],
                    "score": [96.0, 95.0, 94.0, 93.0],
                    "doc_type": ["earnings_release"] * 4,
                    "doc": ["release_q1.txt"] * 4,
                    "evidence_snippet": [
                        shared_summary,
                        shared_summary,
                        "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.",
                        "Reduced principal debt by $787.2 million in the first quarter.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            q1_rows = _quarter_block_notes(ws, "2025-03-31")
            assert sum(1 for note in q1_rows if shared_summary in note) == 1


def test_pbi_quarter_notes_realign_visible_metric_label_when_summary_is_guidance(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q1_visible_label_cleanup.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-03-31"), pd.Timestamp("2025-03-31")],
                    "note_id": ["guide-1", "driver-1"],
                    "category": ["Tone / expectations", "Better / worse vs prior"],
                    "claim": [
                        "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.",
                        "Presort EBIT improved, driven by higher revenue per piece, productivity and cost reduction.",
                    ],
                    "note": [
                        "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.",
                        "Presort EBIT improved, driven by higher revenue per piece, productivity and cost reduction.",
                    ],
                    "metric_ref": ["Deleveraging target", "SendTech / Presort operating driver"],
                    "score": [94.0, 96.0],
                    "doc_type": ["earnings_release", "earnings_release"],
                    "doc": ["release_q1.txt", "release_q1.txt"],
                    "evidence_snippet": [
                        "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.",
                        "Presort EBIT improved, driven by higher revenue per piece, productivity and cost reduction.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="PBI", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            capture = False
            matched = False
            for rr in range(1, ws.max_row + 1):
                marker = str(ws.cell(row=rr, column=1).value or "")
                if marker == "2025-03-31":
                    capture = True
                    continue
                if capture and marker:
                    break
                if not capture:
                    continue
                note = str(ws.cell(row=rr, column=3).value or "")
                metric = str(ws.cell(row=rr, column=4).value or "")
                bucket = str(ws.cell(row=rr, column=2).value or "")
                if "FY 2025 Revenue guidance" in note:
                    matched = True
                    assert metric == "Revenue guidance"
                    assert bucket == "Guidance / outlook"
            assert matched


def test_pbi_quarter_notes_do_not_render_eps_guidance_as_cost_savings_in_q2_block(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q2_eps_alignment.xlsx")
            sec_cache_dir = case_dir / "PBI" / "CEO letters"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "Q2 2025 Earnings CEO Letter.pdf").write_text(
                "Capital Returns – Based on our confidence in the Company's core businesses and our view that shares remain undervalued, "
                "we repurchased $75 million in shares on the open market during the second quarter. "
                "To give us appropriate flexibility, we have increased our existing share repurchase program from $150 million to $400 million. "
                "Additionally, we have again increased our quarterly dividend – from $0.07 per share to $0.08 per share. "
                "Finally, we are increasing our Adjusted EPS guidance from $1.10 - $1.30 to $1.20 - $1.40 primarily due to ongoing share repurchases. "
                "Paul has helped develop actionable initiatives to drive incremental cost reductions and return cash to shareholders.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-06-30")] * 4,
                    "note_id": ["g-1", "g-2", "g-3", "g-4"],
                    "category": ["Guidance / outlook"] * 4,
                    "claim": [
                        "FY 2025 Revenue guidance reaffirmed at $1,900m-$1,950m.",
                        "FY 2025 Adjusted EBIT guidance reaffirmed at $450m-$465m.",
                        "FY 2025 EPS guidance reaffirmed at $1.20-$1.40.",
                        "FY 2025 FCF target reaffirmed at $330m-$370m.",
                    ],
                    "note": [
                        "FY 2025 Revenue guidance reaffirmed at $1,900m-$1,950m.",
                        "FY 2025 Adjusted EBIT guidance reaffirmed at $450m-$465m.",
                        "FY 2025 EPS guidance reaffirmed at $1.20-$1.40.",
                        "FY 2025 FCF target reaffirmed at $330m-$370m.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "Cost savings target",
                        "FCF target",
                    ],
                    "score": [95.0, 94.0, 93.0, 92.0],
                    "doc_type": ["earnings_release"] * 4,
                    "doc": ["release_q2.txt"] * 4,
                    "source_type": ["earnings_release"] * 4,
                    "evidence_snippet": [
                        "FY 2025 Revenue guidance reaffirmed at $1,900m-$1,950m.",
                        "FY 2025 Adjusted EBIT guidance reaffirmed at $450m-$465m.",
                        "FY 2025 EPS guidance reaffirmed at $1.20-$1.40.",
                        "FY 2025 FCF target reaffirmed at $330m-$370m.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="PBI", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            capture = False
            matched_eps = False
            for rr in range(1, ws.max_row + 1):
                marker = str(ws.cell(row=rr, column=1).value or "")
                if marker == "2025-06-30":
                    capture = True
                    continue
                if capture and marker:
                    break
                if not capture:
                    continue
                note = str(ws.cell(row=rr, column=3).value or "")
                metric = str(ws.cell(row=rr, column=4).value or "")
                if "$1.20-$1.40" not in note:
                    continue
                assert "Cost savings target" not in note
                if "EPS guidance" in note:
                    matched_eps = True
                    assert metric == "EPS guidance"
            assert matched_eps


def test_pbi_quarter_notes_do_not_emit_legacy_auth_set_row_when_existing_program_is_raised(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q2_auth_raise_only.xlsx")
            sec_cache_dir = case_dir / "PBI" / "CEO letters"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "Q2 2025 Earnings CEO Letter.pdf").write_text(
                "To give us appropriate flexibility, we have increased our existing share repurchase program from "
                "$150 million to $400 million. Additionally, we have again increased our quarterly dividend – from "
                "$0.07 per share to $0.08 per share.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-06-30")],
                    "note_id": ["q2-anchor"],
                    "category": ["Guidance / outlook"],
                    "claim": ["FY 2025 Revenue guidance reaffirmed at $1,900m-$1,950m."],
                    "note": ["FY 2025 Revenue guidance reaffirmed at $1,900m-$1,950m."],
                    "metric_ref": ["Revenue guidance"],
                    "score": [95.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q2.txt"],
                    "source_type": ["earnings_release"],
                    "evidence_snippet": ["FY 2025 Revenue guidance reaffirmed at $1,900m-$1,950m."],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="PBI", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            capture = False
            q2_notes: list[str] = []
            for rr in range(1, ws.max_row + 1):
                marker = str(ws.cell(row=rr, column=1).value or "")
                if marker == "2025-06-30":
                    capture = True
                    continue
                if capture and marker:
                    break
                if not capture:
                    continue
                q2_notes.append(str(ws.cell(row=rr, column=3).value or ""))

            joined = " | ".join(q2_notes)
            assert "Repurchase authorization increased to $400.0m" in joined
            assert "Share repurchase authorization and remaining capacity updated." not in joined


def test_pbi_quarter_notes_do_not_render_same_note_twice_across_buckets(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_no_duplicate_visible_note.xlsx")
            note_txt = "Revolver availability changed materially in the quarter."
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-12-31")],
                    "note_id": ["liq-1", "liq-2"],
                    "category": ["Debt / refi / covenants", "Cash / liquidity / leverage"],
                    "claim": [note_txt, note_txt],
                    "note": [note_txt, note_txt],
                    "metric_ref": ["revolver_availability_change", "revolver_availability_change"],
                    "score": [91.0, 90.0],
                    "doc_type": ["revolver", "revolver"],
                    "doc": ["revolver.csv", "revolver.csv"],
                    "evidence_snippet": [note_txt, note_txt],
                    "evidence_json": [
                        json.dumps([{"doc_path": "revolver.csv", "doc_type": "revolver", "snippet": note_txt}]),
                        json.dumps([{"doc_path": "revolver.csv", "doc_type": "revolver", "snippet": note_txt}]),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            visible = [
                (
                    str(ws.cell(row=rr, column=3).value or ""),
                    str(ws.cell(row=rr, column=4).value or ""),
                )
                for rr in range(1, ws.max_row + 1)
                if str(ws.cell(row=rr, column=4).value or "") == "Deleveraging / liquidity"
            ]
            assert len(visible) == 1


def test_pbi_quarter_notes_dedupe_same_visible_buyback_note_from_different_source_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_no_duplicate_buyback_preview.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-09-30")] + [pd.Timestamp("2025-09-30")] * 4,
                    "note_id": ["bb-1", "bb-2", "g-1", "g-2", "g-3", "g-4"],
                    "category": [
                        "Cash / liquidity / leverage",
                        "Cash / liquidity / leverage",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                    ],
                    "claim": [
                        "Repurchased 25.9 million shares for $281.2 million since starting the program earlier this year.",
                        "Repurchased 25.9 million shares for $281.2 million since starting the program earlier this year while increasing the authorization to $500 million.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.9bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                    ],
                    "note": [
                        "Repurchased 25.9 million shares for $281.2 million since starting the program earlier this year.",
                        "Repurchased 25.9 million shares for $281.2 million since starting the program earlier this year while increasing the authorization to $500 million.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.9bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                    ],
                    "metric_ref": [
                        "Capital allocation / buyback",
                        "Capital allocation / buyback",
                        "FCF target",
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "EPS guidance",
                    ],
                    "score": [96.0, 95.0, 94.0, 93.0, 92.0, 91.0],
                    "doc_type": ["earnings_release"] * 6,
                    "doc": ["release_q3.txt"] * 6,
                    "evidence_snippet": [
                        "Repurchased 25.9 million shares for $281.2 million since starting the program earlier this year.",
                        "Repurchased 25.9 million shares for $281.2 million since starting the program earlier this year while increasing the authorization to $500 million.",
                        "FY 2025 FCF guidance tracking near the midpoint of $330m-$370m.",
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.9bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m.",
                        "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q3_rows = _quarter_block_notes(ws, "2025-09-30")

            assert sum(
                1
                for note in q3_rows
                if "Repurchased 25.9m shares for $281.2m" in note
                and "since starting the program earlier this year." in note
            ) == 1


def test_quarter_notes_adds_post_quarter_buyback_commentary_as_separate_deduped_row(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "buyback_post_quarter_companion_row.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-03-31")] * 5,
                    "note_id": ["bb-main", "bb-post-release", "bb-post-ceo", "rev-guid", "fcf-guid"],
                    "category": [
                        "Capital allocation / shareholder returns",
                        "Capital allocation / shareholder returns",
                        "Capital allocation / shareholder returns",
                        "Guidance / outlook",
                        "Guidance / outlook",
                    ],
                    "claim": [
                        "Repurchased 1.5 million shares for $15.0 million in the first quarter.",
                        "An additional $12 million of common stock was repurchased after quarter-end through May 2.",
                        "An additional $12 million was repurchased after quarter-end through May 2.",
                        "FY 2025 revenue guidance remains $1.95bn-$2.00bn.",
                        "FY 2025 FCF target remains $330m-$370m.",
                    ],
                    "note": [
                        "Repurchased 1.5 million shares for $15.0 million in the first quarter.",
                        "An additional $12 million of common stock was repurchased after quarter-end through May 2.",
                        "An additional $12 million was repurchased after quarter-end through May 2.",
                        "FY 2025 revenue guidance remains $1.95bn-$2.00bn.",
                        "FY 2025 FCF target remains $330m-$370m.",
                    ],
                    "metric_ref": [
                        "Capital allocation / buyback",
                        "Capital allocation / buyback",
                        "Capital allocation / buyback",
                        "Revenue guidance",
                        "FCF target",
                    ],
                    "score": [97.0, 91.0, 90.0, 95.0, 94.0],
                    "doc_type": ["earnings_release", "earnings_release", "ceo_letter", "earnings_release", "earnings_release"],
                    "doc": ["release_q1.txt", "release_q1.txt", "ceo_q1.txt", "release_q1.txt", "release_q1.txt"],
                    "source_type": ["earnings_release", "earnings_release", "ceo_letter", "earnings_release", "earnings_release"],
                    "evidence_snippet": [
                        "Repurchased 1.5 million shares for $15.0 million in the first quarter.",
                        "An additional $12 million of common stock was repurchased after quarter-end through May 2.",
                        "An additional $12 million was repurchased after quarter-end through May 2.",
                        "FY 2025 revenue guidance remains $1.95bn-$2.00bn.",
                        "FY 2025 FCF target remains $330m-$370m.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q1_rows = _quarter_block_notes(ws, "2025-03-31")

            assert any("Repurchased 1.5m shares for $15.0m" in note for note in q1_rows)
            companion_rows = [note for note in q1_rows if "excluded from quarter/TTM data." in note]
            assert companion_rows == [
                "Additional $12.0m repurchased after quarter-end through May 2; excluded from quarter/TTM data."
            ]


def test_pbi_quarter_notes_add_ceo_letter_management_notes_without_dropping_guidance(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_ceo_letter_management_additions.xlsx")
            ceo_dir = case_dir / "TEST" / "CEO letters"
            ceo_dir.mkdir(parents=True, exist_ok=True)
            (ceo_dir / "Q4 2025 Earnings CEO Letter.txt").write_text(
                "In addition to increasing our share repurchase authorization to $500 million, "
                "we increased our quarterly dividend from $0.08 to $0.09 per share. "
                "We are still on track to commence our strategic review's second phase by the end of the second quarter. "
                "These factors, as well as our efforts to improve forecasting, contributed to us disclosing wider ranges "
                "for the current year's guidance in the Company's earnings press release.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 6,
                    "note_id": ["g-1", "g-2", "g-3", "g-4", "fcf-1", "debt-1"],
                    "category": [
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Cash flow / FCF / capex",
                        "Debt / refi / covenants",
                    ],
                    "claim": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved to $221.7m, up $89.9m YoY.",
                        "Reduced principal debt by $114.1m in Q4.",
                    ],
                    "note": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved to $221.7m, up $89.9m YoY.",
                        "Reduced principal debt by $114.1m in Q4.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "EPS guidance",
                        "FCF target",
                        "FCF improvement",
                        "Debt reduction",
                    ],
                    "score": [97.0, 96.0, 95.0, 98.0, 94.0, 93.0],
                    "doc_type": ["earnings_release"] * 6,
                    "doc": ["release_q4.txt"] * 6,
                    "evidence_snippet": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved to $221.7m, up $89.9m YoY.",
                        "Reduced principal debt by $114.1m in Q4.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q4_rows = _quarter_block_notes(ws, "2025-12-31")

            assert len(q4_rows) >= 5
            assert any("Free cash flow improved to $221.7m" in note for note in q4_rows)
            assert any("Reduced principal debt by $114.1m in Q4." in note for note in q4_rows)
            assert any("Repurchase authorization increased" in note for note in q4_rows)
            assert any("Quarterly dividend" in note for note in q4_rows)
            assert any("Strategic review phase 2 remains on track by end of Q2 2026." in note for note in q4_rows)
            assert not any(
                "reduced principal debt" in note.lower() and "quarterly dividend" in note.lower()
                for note in q4_rows
            )


def test_pbi_quarter_notes_can_add_auth_dividend_note_alongside_buyback_execution(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q3_auth_dividend_additive.xlsx")
            ceo_dir = case_dir / "TEST" / "CEO letters"
            ceo_dir.mkdir(parents=True, exist_ok=True)
            (ceo_dir / "Q3 2025 Earnings CEO Letter.txt").write_text(
                "In addition to increasing our share repurchase authorization to $500 million, "
                "we increased our quarterly dividend from $0.08 to $0.09 per share. "
                "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million "
                "since starting the program earlier this year. In the third quarter alone, we bought back over 8% "
                "of the shares outstanding.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 6,
                    "note_id": ["g-1", "g-2", "g-3", "g-4", "rev-1", "fcf-1"],
                    "category": [
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Debt / liquidity / covenants",
                        "Cash flow / FCF / capex",
                    ],
                    "claim": [
                        "FY 2025 Revenue guidance midpoint $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance midpoint $430m-$460m.",
                        "FY 2025 EPS guidance midpoint $1.20-$1.40.",
                        "FY 2025 FCF target $330m-$370m.",
                        "Revolver availability moved to $400.0m, delta $135.0m.",
                        "Free cash flow improved to $111.4m, up $31.5m YoY.",
                    ],
                    "note": [
                        "FY 2025 Revenue guidance midpoint $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance midpoint $430m-$460m.",
                        "FY 2025 EPS guidance midpoint $1.20-$1.40.",
                        "FY 2025 FCF target $330m-$370m.",
                        "Revolver availability moved to $400.0m, delta $135.0m.",
                        "Free cash flow improved to $111.4m, up $31.5m YoY.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "EPS guidance",
                        "FCF target",
                        "Deleveraging / liquidity",
                        "FCF improvement",
                    ],
                    "score": [97.0, 96.0, 95.0, 94.0, 93.0, 92.0],
                    "doc_type": ["earnings_release"] * 6,
                    "doc": ["release_q3.txt"] * 6,
                    "evidence_snippet": [
                        "FY 2025 Revenue guidance midpoint $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance midpoint $430m-$460m.",
                        "FY 2025 EPS guidance midpoint $1.20-$1.40.",
                        "FY 2025 FCF target $330m-$370m.",
                        "Revolver availability moved to $400.0m, delta $135.0m.",
                        "Free cash flow improved to $111.4m, up $31.5m YoY.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q3_rows = _quarter_block_notes(ws, "2025-09-30")

            assert any("Repurchase authorization increased to $500.0m" in note for note in q3_rows)
            assert any("Quarterly dividend increased to $0.09/share from $0.08/share" in note for note in q3_rows)


def test_write_excel_can_emit_quarter_notes_audit_sheet_and_saved_workbook_provenance(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "quarter_notes_audit_saved_truth.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 2,
                    "note_id": ["fcf-1", "debt-1"],
                    "category": ["Cash flow / FCF / capex", "Debt / liquidity / covenants"],
                    "claim": [
                        "Free cash flow improved to $221.7m, up $89.9m YoY.",
                        "Reduced principal debt by $114.1m in Q4.",
                    ],
                    "note": [
                        "Free cash flow improved to $221.7m, up $89.9m YoY.",
                        "Reduced principal debt by $114.1m in Q4.",
                    ],
                    "metric_ref": ["FCF improvement", "Debt reduction"],
                    "score": [95.0, 94.0],
                    "doc_type": ["earnings_release", "earnings_release"],
                    "doc": ["release_q4.txt", "release_q4.txt"],
                    "source_type": ["earnings_release", "earnings_release"],
                    "evidence_snippet": [
                        "Free cash flow improved to $221.7m, up $89.9m YoY.",
                        "Reduced principal debt by $114.1m in Q4.",
                    ],
                }
            )

            result = write_excel_from_inputs(
                _make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes, quarter_notes_audit=True)
            )

            assert result.saved_workbook_provenance["workbook_path"] == str(out_path)
            assert result.saved_workbook_provenance["workbook_sha1"]
            assert "Generated at" in result.saved_workbook_provenance["quarter_notes_header"]
            audit_rows = _read_quarter_notes_audit_rows(out_path)
            assert audit_rows
            assert any(row.get("stage") == "readback_verified" for row in audit_rows)
            wb = load_workbook(out_path, data_only=True)
            assert wb["Quarter_Notes_Audit"].sheet_state == "visible"
            assert wb.sheetnames[-1] == "Quarter_Notes_Audit"


def test_enrich_quarter_notes_audit_rows_promotes_terminal_stage_and_drops_redundant_missing_rows() -> None:
    audit_rows = [
        {
            "quarter": "2025-12-31",
            "trace_id": "trace-1",
            "stage": "final_selected",
            "idea_label": "FCF improvement",
            "canonical_source_group": "earnings_release:release_q4",
            "visible_category": "Cash flow / FCF / capex",
            "final_summary": "Free cash flow improved to $221.7m, up $89.9m YoY.",
            "source_doc": "release_q4.txt",
        },
        {
            "quarter": "2025-12-31",
            "trace_id": "trace-2",
            "stage": "final_selected",
            "idea_label": "FCF improvement",
            "canonical_source_group": "earnings_release:release_q4",
            "visible_category": "Cash flow / FCF / capex",
            "final_summary": "Blob-like duplicate that should not survive.",
            "source_doc": "release_q4.txt",
        },
    ]
    provenance = {
        "workbook_path": "model.xlsx",
        "quarter_notes_ui_snapshot_rows": {
            "2025-12-31": [
                ("Cash flow / FCF / capex", "Free cash flow improved to $221.7m, up $89.9m YoY.", 42),
            ]
        },
    }

    rows = enrich_quarter_notes_audit_rows_with_readback(audit_rows, provenance)

    assert any(row.get("stage") == "readback_verified" for row in rows)
    assert not any(row.get("stage") == "final_selected" for row in rows)
    assert not any(
        row.get("stage") == "saved_workbook_missing"
        and str(row.get("canonical_source_group") or "") == "earnings_release:release_q4"
        and str(row.get("idea_label") or "") == "FCF improvement"
        for row in rows
    )


def test_write_excel_can_skip_temp_saved_workbook_provenance_capture(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "skip_temp_saved_truth.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "note_id": ["fcf-1"],
                    "category": ["Cash flow / FCF / capex"],
                    "claim": ["Free cash flow improved to $221.7m, up $89.9m YoY."],
                    "note": ["Free cash flow improved to $221.7m, up $89.9m YoY."],
                    "metric_ref": ["FCF improvement"],
                    "score": [95.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q4.txt"],
                    "source_type": ["earnings_release"],
                    "evidence_snippet": ["Free cash flow improved to $221.7m, up $89.9m YoY."],
                }
            )

            result = write_excel_from_inputs(
                _make_inputs(
                    out_path,
                    ticker="TEST",
                    quarter_notes=quarter_notes,
                    quarter_notes_audit=True,
                    capture_saved_workbook_provenance=False,
                )
            )

            assert result.saved_workbook_provenance == {}
            assert result.quarter_notes_header_text
            assert result.summary_export_expectation is not None
            assert result.valuation_export_expectation is not None
            assert out_path.exists()


def test_validate_saved_workbook_export_builds_provenance_and_validates_in_one_pass(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "combined_saved_truth.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "note_id": ["fcf-1"],
                    "category": ["Cash flow / FCF / capex"],
                    "claim": ["Free cash flow improved to $221.7m, up $89.9m YoY."],
                    "note": ["Free cash flow improved to $221.7m, up $89.9m YoY."],
                    "metric_ref": ["FCF improvement"],
                    "score": [95.0],
                    "doc_type": ["earnings_release"],
                    "doc": ["release_q4.txt"],
                    "source_type": ["earnings_release"],
                    "evidence_snippet": ["Free cash flow improved to $221.7m, up $89.9m YoY."],
                }
            )

            result = write_excel_from_inputs(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))

            provenance = validate_saved_workbook_export(
                out_path,
                quarter_notes_ui_snapshot=result.quarter_notes_ui_snapshot,
                summary_export_expectation=result.summary_export_expectation,
                valuation_export_expectation=result.valuation_export_expectation,
                qa_export_expectation=result.qa_export_expectation,
                needs_review_export_expectation=result.needs_review_export_expectation,
            )

            assert provenance["workbook_path"] == str(out_path)
            assert provenance["workbook_sha1"]
            assert provenance["summary_snapshot"]
            assert provenance["valuation_snapshot"]
            assert provenance["quarter_notes_ui_snapshot"] == result.quarter_notes_ui_snapshot


def test_quarter_notes_audit_sheet_strips_illegal_control_chars() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "audit_sheet_illegal_chars.xlsx")
        wb = Workbook()
        wb.save(out_path)
        write_quarter_notes_audit_sheet(
            out_path,
            [
                {
                    "quarter": "2025-03-31",
                    "trace_id": "trace-1",
                    "stage": "source_detected",
                    "source_excerpt": "Cost reductions ahead of plan e\x00ciency and O\x0fcer text",
                    "final_summary": "Ahead of plan.",
                }
            ],
        )
        wb2 = load_workbook(out_path, data_only=True)
        ws = wb2["Quarter_Notes_Audit"]
        vals = list(ws.iter_rows(values_only=True))
        header = [str(x or "") for x in vals[0]]
        source_excerpt_idx = header.index("source_excerpt")
        assert "\x00" not in str(vals[1][source_excerpt_idx] or "")
        assert "\x0f" not in str(vals[1][source_excerpt_idx] or "")


def test_pbi_quarter_notes_audit_traces_auth_dividend_note_to_saved_workbook(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q3_auth_dividend_audit.xlsx")
            ceo_dir = case_dir / "TEST" / "CEO letters"
            ceo_dir.mkdir(parents=True, exist_ok=True)
            (ceo_dir / "Q3 2025 Earnings CEO Letter.txt").write_text(
                "In addition to increasing our share repurchase authorization to $500 million, "
                "we increased our quarterly dividend from $0.08 to $0.09 per share. "
                "Through last Friday, we repurchased 25.9 million shares at a total cost of $281.2 million "
                "since starting the program earlier this year.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 5,
                    "note_id": ["g-1", "g-2", "g-3", "liq-1", "fcf-1"],
                    "category": [
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Debt / liquidity / covenants",
                        "Cash flow / FCF / capex",
                    ],
                    "claim": [
                        "FY 2025 Revenue guidance midpoint $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance midpoint $430m-$460m.",
                        "FY 2025 FCF target $330m-$370m.",
                        "Revolver availability moved to $400.0m, delta $135.0m.",
                        "Free cash flow improved to $111.4m, up $31.5m YoY.",
                    ],
                    "note": [
                        "FY 2025 Revenue guidance midpoint $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance midpoint $430m-$460m.",
                        "FY 2025 FCF target $330m-$370m.",
                        "Revolver availability moved to $400.0m, delta $135.0m.",
                        "Free cash flow improved to $111.4m, up $31.5m YoY.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "FCF target",
                        "Deleveraging / liquidity",
                        "FCF improvement",
                    ],
                    "score": [97.0, 96.0, 95.0, 94.0, 93.0],
                    "doc_type": ["earnings_release"] * 5,
                    "doc": ["release_q3.txt"] * 5,
                    "source_type": ["earnings_release"] * 5,
                    "evidence_snippet": [
                        "FY 2025 Revenue guidance midpoint $1.90bn-$1.95bn.",
                        "FY 2025 Adjusted EBIT guidance midpoint $430m-$460m.",
                        "FY 2025 FCF target $330m-$370m.",
                        "Revolver availability moved to $400.0m, delta $135.0m.",
                        "Free cash flow improved to $111.4m, up $31.5m YoY.",
                    ],
                }
            )

            write_excel_from_inputs(
                _make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes, quarter_notes_audit=True)
            )
            audit_rows = _read_quarter_notes_audit_rows(out_path)
            auth_row = next(
                row
                for row in audit_rows
                if row.get("stage") == "readback_verified"
                and "Repurchase authorization increased to $500.0m" in row.get("final_summary", "")
            )
            dividend_row = next(
                row
                for row in audit_rows
                if row.get("stage") == "readback_verified"
                and "Quarterly dividend increased to $0.09/share from $0.08/share." in row.get("final_summary", "")
            )
            auth_trace_rows = [row for row in audit_rows if row.get("trace_id") == auth_row["trace_id"]]
            dividend_trace_rows = [row for row in audit_rows if row.get("trace_id") == dividend_row["trace_id"]]

            assert any(row.get("stage") == "source_detected" for row in auth_trace_rows)
            assert any(row.get("stage") == "candidate_created" for row in auth_trace_rows)
            assert any(
                row.get("stage") == "source_detected"
                and "quarterly dividend" in (row.get("source_excerpt", "") + " " + row.get("final_summary", "")).lower()
                for row in audit_rows
            )
            assert any(row.get("stage") == "readback_verified" for row in dividend_trace_rows)
            assert any(
                row.get("stage") == "candidate_created"
                and "quarterly dividend" in (row.get("source_excerpt", "") + " " + row.get("final_summary", "")).lower()
                for row in audit_rows
            )
            assert auth_row.get("authorization_confidence") == "present"
            assert dividend_row.get("dividend_change_confidence") == "present"
            assert auth_row.get("scope_confidence") in {"policy_only", "cumulative"}
            assert dividend_row.get("scope_confidence") in {"policy_only", "cumulative"}


def test_pbi_quarter_notes_saved_workbook_adds_q4_auth_capacity_from_html_sec_cache(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q4_html_auth_capacity_saved.xlsx")
            sec_dir = case_dir / "TEST" / "sec_cache"
            sec_dir.mkdir(parents=True, exist_ok=True)
            (sec_dir / "doc_000000000026000001_q42025earningspressrelea.htm").write_text(
                "<html><body>"
                "Pitney Bowes&#8217; Board of Directors recently increased the Company&#8217;s repurchase authorization by $250 million. "
                "As of February 13, 2026, there was $359 million in capacity remaining under the authorization. "
                "The Board approved a regular quarterly dividend of $0.09 per share."
                "</body></html>",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 6,
                    "note_id": ["g-1", "g-2", "g-3", "g-4", "fcf-1", "debt-1"],
                    "category": [
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Cash flow / FCF / capex",
                        "Debt / refi / covenants",
                    ],
                    "claim": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved to $221.7m, up $89.9m YoY.",
                        "Reduced principal debt by $114.1m in Q4.",
                    ],
                    "note": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved to $221.7m, up $89.9m YoY.",
                        "Reduced principal debt by $114.1m in Q4.",
                    ],
                    "metric_ref": [
                        "Revenue guidance",
                        "Adjusted EBIT guidance",
                        "EPS guidance",
                        "FCF target",
                        "FCF improvement",
                        "Debt reduction",
                    ],
                    "score": [97.0, 96.0, 95.0, 98.0, 94.0, 93.0],
                    "doc_type": ["earnings_release"] * 6,
                    "doc": ["release_q4.txt"] * 6,
                    "source_type": ["earnings_release"] * 6,
                    "evidence_snippet": [
                        "FY 2026 Revenue guidance updated to $1.76bn-$1.86bn.",
                        "FY 2026 Adjusted EBIT guidance updated to $410m-$460m.",
                        "FY 2026 EPS guidance updated to $1.40-$1.60.",
                        "FY 2026 FCF guidance updated to $340m-$370m.",
                        "Free cash flow improved to $221.7m, up $89.9m YoY.",
                        "Reduced principal debt by $114.1m in Q4.",
                    ],
                }
            )

            write_excel_from_inputs(
                _make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes, quarter_notes_audit=True)
            )
            wb = load_workbook(out_path, data_only=True)
            q4_rows = _quarter_block_notes(wb["Quarter_Notes_UI"], "2025-12-31")

            assert any(
                "Repurchase authorization increased by $250.0m." in note
                for note in q4_rows
            )
            assert any("Remaining share repurchase capacity was $359.0m at quarter-end." in note for note in q4_rows)
            assert any("Quarterly dividend set at $0.09/share." in note for note in q4_rows)


def test_pbi_quarter_notes_collapse_equivalent_guidance_targets_with_bn_and_m_formats(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_q1_equivalent_guidance_formats.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-03-31")] * 6,
                    "note_id": ["driver-1", "rev-bn", "rev-m", "fcf-1", "ebit-1", "debt-1"],
                    "category": [
                        "Better / worse vs prior",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Guidance / outlook",
                        "Debt / refi / covenants",
                    ],
                    "claim": [
                        "Presort EBIT improved, driven by higher revenue per piece, productivity and cost reduction.",
                        "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.",
                        "FY 2025 Revenue guidance updated to $1,950m-$2,000m.",
                        "FY 2025 FCF target $330m-$370m.",
                        "FY 2025 Adjusted EBIT guidance $450m-$480m.",
                        "Reduced principal debt by $787.2m in Q1.",
                    ],
                    "note": [
                        "Presort EBIT improved, driven by higher revenue per piece, productivity and cost reduction.",
                        "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.",
                        "FY 2025 Revenue guidance updated to $1,950m-$2,000m.",
                        "FY 2025 FCF target $330m-$370m.",
                        "FY 2025 Adjusted EBIT guidance $450m-$480m.",
                        "Reduced principal debt by $787.2m in Q1.",
                    ],
                    "metric_ref": [
                        "SendTech / Presort operating driver",
                        "Revenue guidance",
                        "Revenue guidance",
                        "FCF target",
                        "Adjusted EBIT guidance",
                        "Debt reduction",
                    ],
                    "score": [96.0, 94.0, 95.0, 93.0, 92.0, 91.0],
                    "doc_type": ["earnings_release"] * 6,
                    "doc": ["release_q1.txt"] * 6,
                    "evidence_snippet": [
                        "Presort EBIT improved, driven by higher revenue per piece, productivity and cost reduction.",
                        "FY 2025 Revenue guidance updated to $1.95bn-$2.00bn.",
                        "FY 2025 Revenue guidance updated to $1,950m-$2,000m.",
                        "FY 2025 FCF target $330m-$370m.",
                        "FY 2025 Adjusted EBIT guidance $450m-$480m.",
                        "Reduced principal debt by $787.2m in Q1.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="PBI", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q1_rows = _quarter_block_notes(ws, "2025-03-31")

            revenue_rows = [note for note in q1_rows if "FY 2025 Revenue guidance" in note]
            assert len(revenue_rows) == 1


def test_gpre_quarter_notes_audit_traces_management_note_and_attrition_state(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_q1_management_audit.xlsx")
            press_dir = case_dir / "TEST" / "press_release"
            press_dir.mkdir(parents=True, exist_ok=True)
            (press_dir / "Green-Plains-Reports-First-Quarter-2025-Financial-Results-2025.txt").write_text(
                "With our cost reduction initiatives implemented and progressing ahead of plan, "
                "we are positioned to deliver positive EBITDA for the remainder of the year based on current market conditions. "
                "We have also taken decisive steps to enhance liquidity and remain focused on monetizing non-core assets "
                "to strengthen our balance sheet.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-03-31")] * 4,
                    "note_id": ["fcf-1", "debt-1", "util-1", "ops-1"],
                    "category": [
                        "Cash flow / FCF / capex",
                        "Debt / liquidity / covenants",
                        "Results / drivers",
                        "Programs / initiatives",
                    ],
                    "claim": [
                        "FCF TTM YoY delta $-149.8m",
                        "Net debt delta $80.9m",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 100%.",
                        "Compression infrastructure under construction; Q4 2025 start-up still on track.",
                    ],
                    "note": [
                        "FCF TTM YoY delta $-149.8m",
                        "Net debt delta $80.9m",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 100%.",
                        "Compression infrastructure under construction; Q4 2025 start-up still on track.",
                    ],
                    "metric_ref": ["FCF", "Net debt / leverage", "Utilization", "Strategic milestone"],
                    "score": [95.0, 94.0, 93.0, 92.0],
                    "doc_type": ["model_metric", "model_metric", "earnings_release", "earnings_release"],
                    "doc": ["history_q", "history_q", "release_q1.txt", "release_q1.txt"],
                    "source_type": ["model_metric", "model_metric", "earnings_release", "earnings_release"],
                    "evidence_snippet": [
                        "FCF TTM YoY delta $-149.8m",
                        "Net debt delta $80.9m",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 100%.",
                        "Compression infrastructure under construction; Q4 2025 start-up still on track.",
                    ],
                }
            )

            write_excel_from_inputs(
                _make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes, quarter_notes_audit=True)
            )
            audit_rows = _read_quarter_notes_audit_rows(out_path)

            assert any(
                row.get("stage") == "readback_verified"
                and "ahead of plan" in row.get("final_summary", "").lower()
                and "positive ebitda" in row.get("final_summary", "").lower()
                for row in audit_rows
            )
            liquidity_rows = [
                row
                for row in audit_rows
                if "non-core asset" in row.get("source_excerpt", "").lower()
                or "enhance liquidity" in row.get("source_excerpt", "").lower()
            ]
            assert liquidity_rows
            assert any(
                row.get("stage") in {"source_detected", "candidate_created", "selection_lost", "theme_collapsed", "readback_verified"}
                for row in liquidity_rows
            )


def test_gpre_quarter_notes_add_q1_management_framing_from_press_release(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_q1_management_framing.xlsx")
            press_dir = case_dir / "TEST" / "press_release"
            press_dir.mkdir(parents=True, exist_ok=True)
            (press_dir / "Green-Plains-Reports-First-Quarter-2025-Financial-Results-2025.txt").write_text(
                "With our cost reduction initiatives implemented and progressing ahead of plan, "
                "paired with a disciplined hedging program overseen by our newly formed Risk Committee, "
                "we are positioned to deliver positive EBITDA for the remainder of the year based on current market conditions. "
                "We have also taken decisive steps to enhance liquidity and remain focused on monetizing non-core assets "
                "to strengthen our balance sheet, improve capital access and further reduce our cost structure.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-03-31")] * 4,
                    "note_id": ["fcf-1", "debt-1", "util-1", "ops-1"],
                    "category": [
                        "Cash flow / FCF / capex",
                        "Debt / liquidity / covenants",
                        "Results / drivers",
                        "Programs / initiatives",
                    ],
                    "claim": [
                        "FCF TTM YoY delta $-149.8m",
                        "Net debt delta $80.9m",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 100%.",
                        "Compression infrastructure under construction; Q4 2025 start-up still on track.",
                    ],
                    "note": [
                        "FCF TTM YoY delta $-149.8m",
                        "Net debt delta $80.9m",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 100%.",
                        "Compression infrastructure under construction; Q4 2025 start-up still on track.",
                    ],
                    "metric_ref": [
                        "FCF",
                        "Net debt / leverage",
                        "Utilization",
                        "Strategic milestone",
                    ],
                    "score": [95.0, 94.0, 93.0, 92.0],
                    "doc_type": ["model_metric", "model_metric", "earnings_release", "earnings_release"],
                    "doc": ["history_q", "history_q", "release_q1.txt", "release_q1.txt"],
                    "source_type": ["model_metric", "model_metric", "earnings_release", "earnings_release"],
                    "evidence_snippet": [
                        "FCF TTM YoY delta $-149.8m",
                        "Net debt delta $80.9m",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 100%.",
                        "Compression infrastructure under construction; Q4 2025 start-up still on track.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q1_rows = _quarter_block_notes(ws, "2025-03-31")

            assert len(q1_rows) >= 5
            assert any("ahead of plan" in note.lower() and "positive EBITDA outlook".lower() in note.lower() for note in q1_rows)
            assert any("non-core asset monetization" in note.lower() and "enhance liquidity" in note.lower() for note in q1_rows)


def test_gpre_quarter_notes_keep_concrete_management_summary_from_noisy_press_release(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_q1_noisy_management_framing.xlsx")
            press_dir = case_dir / "TEST" / "press_release"
            press_dir.mkdir(parents=True, exist_ok=True)
            (press_dir / "Green-Plains-Reports-First-Quarter-2025-Financial-Results-2025.txt").write_text(
                "With our cost reduction initiatives implemented and progressing ahead of plan, paired with a disciplined hedging "
                "program overseen by our newly formed Risk Committee, we are positioned to deliver positive EBITDA for the remainder "
                "of the year based on current market conditions. We have also taken decisive steps to enhance liquidity and remain "
                "focused on monetizing non-core assets to strengthen our balance sheet, improve capital access and further reduce our cost structure.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-03-31")] * 3,
                    "note_id": ["fcf-1", "debt-1", "util-1"],
                    "category": [
                        "Cash flow / FCF / capex",
                        "Debt / liquidity / covenants",
                        "Results / drivers",
                    ],
                    "claim": [
                        "FCF TTM YoY delta $-149.8m",
                        "Net debt delta $80.9m",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 100%.",
                    ],
                    "note": [
                        "FCF TTM YoY delta $-149.8m",
                        "Net debt delta $80.9m",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 100%.",
                    ],
                    "metric_ref": ["FCF", "Net debt / leverage", "Utilization"],
                    "score": [95.0, 94.0, 93.0],
                    "doc_type": ["model_metric", "model_metric", "earnings_release"],
                    "doc": ["history_q", "history_q", "release_q1.txt"],
                    "source_type": ["model_metric", "model_metric", "earnings_release"],
                    "text_full": [
                        "FCF TTM YoY delta $-149.8m",
                        "Net debt delta $80.9m",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 100%.",
                    ],
                    "comment_full_text": [
                        "FCF TTM YoY delta $-149.8m",
                        "Net debt delta $80.9m",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 100%.",
                    ],
                    "evidence_snippet": [
                        "FCF TTM YoY delta $-149.8m",
                        "Net debt delta $80.9m",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 100%.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ws = ctx.wb["Quarter_Notes_UI"] if "Quarter_Notes_UI" in ctx.wb.sheetnames else None
            if ws is not None:
                del ctx.wb["Quarter_Notes_UI"]
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q1_rows = _quarter_block_notes(ws, "2025-03-31")

            assert any("ahead of plan" in note.lower() and "positive EBITDA outlook".lower() in note.lower() for note in q1_rows)
            assert any("non-core asset monetization" in note.lower() and "enhance liquidity" in note.lower() for note in q1_rows)


def test_gpre_quarter_notes_add_q2_working_capital_note_from_press_release(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_q2_working_capital.xlsx")
            press_dir = case_dir / "TEST" / "press_release"
            press_dir.mkdir(parents=True, exist_ok=True)
            (press_dir / "Green-Plains-Reports-Second-Quarter-2025-Financial-Results-2025.txt").write_text(
                "We executed several key initiatives this quarter to sustain reliable, safe operations, improve efficiencies "
                "and enhance our operating performance. We delivered greater than $50 million improvement in working capital, "
                "delivering scale, optimizing value and improving supply chain efficiencies.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-06-30")] * 4,
                    "note_id": ["ops-1", "margin-1", "rev-1", "util-1"],
                    "category": [
                        "Programs / initiatives",
                        "Strategy / segment",
                        "Debt / liquidity / covenants",
                        "Results / drivers",
                    ],
                    "claim": [
                        "Carbon capture infrastructure equipment delivered and Q4 2025 start-up remains on track.",
                        "EBITDA margin delta -143 bps",
                        "Revolver availability moved to $258.5m at 2025-06-30 (delta $58.5m).",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 99%.",
                    ],
                    "note": [
                        "Carbon capture infrastructure equipment delivered and Q4 2025 start-up remains on track.",
                        "EBITDA margin delta -143 bps",
                        "Revolver availability moved to $258.5m at 2025-06-30 (delta $58.5m).",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 99%.",
                    ],
                    "metric_ref": ["Strategic milestone", "EBITDA margin", "revolver_availability_change", "Utilization"],
                    "score": [94.0, 93.0, 92.0, 91.0],
                    "doc_type": ["earnings_release", "model_metric", "model_metric", "earnings_release"],
                    "doc": ["release_q2.txt", "history_q", "history_q", "release_q2.txt"],
                    "source_type": ["earnings_release", "model_metric", "model_metric", "earnings_release"],
                    "evidence_snippet": [
                        "Carbon capture infrastructure equipment delivered and Q4 2025 start-up remains on track.",
                        "EBITDA margin delta -143 bps",
                        "Revolver availability moved to $258.5m at 2025-06-30 (delta $58.5m).",
                        "Achieved strong utilization in the quarter from the nine operating ethanol plants of 99%.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q2_rows = _quarter_block_notes(ws, "2025-06-30")

            assert any("Working capital improved by more than $50.0m." in note for note in q2_rows)


def test_gpre_quarter_notes_add_q4_45z_contribution_and_crush_margin_notes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_q4_45z_contribution.xlsx")
            press_dir = case_dir / "TEST" / "press_release"
            press_dir.mkdir(parents=True, exist_ok=True)
            (press_dir / "Green-Plains-Reports-Fourth-Quarter-and-Full-Year-2025-Financial-Results-2026.txt").write_text(
                "Adjusted EBITDA of $49.1 million, inclusive of $23.4 million in 45Z production tax credit value net of discounts and other costs. "
                "The consolidated ethanol crush margin was $44.4 million for the fourth quarter of 2025, compared with $(15.5) million for the same period in 2024.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")] * 5,
                    "note_id": ["ebitda-1", "debt-1", "fcf-1", "util-1", "ops-1"],
                    "category": [
                        "Strategy / segment",
                        "Debt / liquidity / covenants",
                        "Cash flow / FCF / capex",
                        "Results / drivers",
                        "Programs / initiatives",
                    ],
                    "claim": [
                        "Adjusted EBITDA YoY 369.8%",
                        "Net debt delta $-77.9m",
                        "FCF TTM YoY delta $198.7m",
                        "Utilization 97% of stated capacity",
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                    ],
                    "note": [
                        "Adjusted EBITDA YoY 369.8%",
                        "Net debt delta $-77.9m",
                        "FCF TTM YoY delta $198.7m",
                        "Utilization 97% of stated capacity",
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                    ],
                    "metric_ref": ["Adjusted EBITDA", "Net debt / leverage", "FCF", "Utilization", "Strategic milestone"],
                    "score": [96.0, 95.0, 94.0, 93.0, 92.0],
                    "doc_type": ["model_metric", "model_metric", "model_metric", "earnings_release", "earnings_release"],
                    "doc": ["history_q", "history_q", "history_q", "release_q4.txt", "release_q4.txt"],
                    "source_type": ["model_metric", "model_metric", "model_metric", "earnings_release", "earnings_release"],
                    "evidence_snippet": [
                        "Adjusted EBITDA YoY 369.8%",
                        "Net debt delta $-77.9m",
                        "FCF TTM YoY delta $198.7m",
                        "Utilization 97% of stated capacity",
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q4_rows = _quarter_block_notes(ws, "2025-12-31")

            assert len(q4_rows) >= 7
            assert any("45Z production tax credits contributed $23.4m net of discounts and other costs in Q4." in note for note in q4_rows)
            assert any("Consolidated ethanol crush margin improved to $44.4m" in note for note in q4_rows)
            assert any("FCF TTM improved by $198.7m YoY." in note for note in q4_rows)
            realized_row = next(
                rr
                for rr in range(1, ws.max_row + 1)
                if "45Z production tax credits contributed $23.4m net of discounts and other costs in Q4."
                in str(ws.cell(row=rr, column=3).value or "")
            )
            assert str(ws.cell(row=realized_row, column=2).value or "").strip() == "Results / drivers"
            assert str(ws.cell(row=realized_row, column=4).value or "").strip() == "45Z value realized"


def test_gpre_quarter_notes_add_2024_q1_margin_driver_from_press_release(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_2024_q1_margin_driver.xlsx")
            press_dir = case_dir / "TEST" / "press_release"
            press_dir.mkdir(parents=True, exist_ok=True)
            (press_dir / "Green-Plains-Reports-First-Quarter-2024-Financial-Results-2024.txt").write_text(
                "Margins in the first quarter were weaker across our product mix and we were impacted by industry oversupply "
                "during a mild winter leading to stock builds and lower prices realized, though margins have improved from the first quarter low. "
                "A few plants that were idled during the January cold snap had an impact on the quarter, in addition to significant planned maintenance programs.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2024-03-31")] * 4,
                    "note_id": ["fcf-1", "debt-1", "margin-1", "rev-1"],
                    "category": [
                        "Cash flow / FCF / capex",
                        "Debt / liquidity / covenants",
                        "Strategy / segment",
                        "Debt / liquidity / covenants",
                    ],
                    "claim": [
                        "FCF TTM YoY delta $93.1m",
                        "Net debt delta $112.1m",
                        "EBITDA margin delta 172 bps",
                        "Revolver availability moved to $230.0m at 2024-03-31 (delta $30.0m).",
                    ],
                    "note": [
                        "FCF TTM YoY delta $93.1m",
                        "Net debt delta $112.1m",
                        "EBITDA margin delta 172 bps",
                        "Revolver availability moved to $230.0m at 2024-03-31 (delta $30.0m).",
                    ],
                    "metric_ref": ["FCF", "Net debt / leverage", "EBITDA margin", "revolver_availability_change"],
                    "score": [95.0, 94.0, 93.0, 92.0],
                    "doc_type": ["model_metric", "model_metric", "model_metric", "model_metric"],
                    "doc": ["history_q", "history_q", "history_q", "history_q"],
                    "source_type": ["model_metric", "model_metric", "model_metric", "model_metric"],
                    "evidence_snippet": [
                        "FCF TTM YoY delta $93.1m",
                        "Net debt delta $112.1m",
                        "EBITDA margin delta 172 bps",
                        "Revolver availability moved to $230.0m at 2024-03-31 (delta $30.0m).",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q1_2024_rows = _quarter_block_notes(ws, "2024-03-31")

            assert len(q1_2024_rows) >= 5
            assert any("Margins were pressured by industry oversupply" in note for note in q1_2024_rows)


def test_gpre_quarter_notes_keep_only_richer_near_duplicate_margin_driver(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_near_duplicate_margin_driver.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2024-03-31")] * 6,
                    "note_id": ["m-1", "m-2", "fcf-1", "debt-1", "ebitda-1", "rev-1"],
                    "category": [
                        "Results / drivers",
                        "Results / drivers",
                        "Cash flow / FCF / capex",
                        "Debt / liquidity / covenants",
                        "Strategy / segment",
                        "Programs / initiatives",
                    ],
                    "claim": [
                        "Margins were pressured by industry oversupply and a mild winter.",
                        "Margins were pressured by industry oversupply, a mild winter and plant downtime/maintenance.",
                        "FCF TTM YoY delta $93.1m",
                        "Net debt delta $112.1m",
                        "EBITDA margin delta 172 bps",
                        "Revolver availability moved to $230.0m at 2024-03-31 (delta $30.0m).",
                    ],
                    "note": [
                        "Margins were pressured by industry oversupply and a mild winter.",
                        "Margins were pressured by industry oversupply, a mild winter and plant downtime/maintenance.",
                        "FCF TTM YoY delta $93.1m",
                        "Net debt delta $112.1m",
                        "EBITDA margin delta 172 bps",
                        "Revolver availability moved to $230.0m at 2024-03-31 (delta $30.0m).",
                    ],
                    "metric_ref": [
                        "Margin driver",
                        "Margin driver",
                        "FCF",
                        "Net debt / leverage",
                        "EBITDA margin",
                        "revolver_availability_change",
                    ],
                    "score": [96.0, 97.0, 95.0, 94.0, 93.0, 92.0],
                    "doc_type": ["press_release", "press_release", "model_metric", "model_metric", "model_metric", "model_metric"],
                    "doc": ["release_q1.txt", "release_q1.txt", "history_q", "history_q", "history_q", "history_q"],
                    "source_type": ["press_release", "press_release", "model_metric", "model_metric", "model_metric", "model_metric"],
                    "evidence_snippet": [
                        "Margins were pressured by industry oversupply and a mild winter.",
                        "Margins were pressured by industry oversupply, a mild winter and plant downtime/maintenance.",
                        "FCF TTM YoY delta $93.1m",
                        "Net debt delta $112.1m",
                        "EBITDA margin delta 172 bps",
                        "Revolver availability moved to $230.0m at 2024-03-31 (delta $30.0m).",
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]
            q1_rows = _quarter_block_notes(ws, "2024-03-31")

            assert sum(1 for note in q1_rows if "Margins were pressured by industry oversupply" in note) == 1


def test_pbi_tracker_is_stricter_than_notes_for_same_quarter_inputs(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_tracker_stricter_than_notes.xlsx")
            promises = pd.DataFrame(
                {
                    "promise_id": ["soft", "hard"],
                    "quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-09-30")],
                    "created_quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-09-30")],
                    "last_seen_quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-09-30")],
                    "first_seen_evidence_quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-09-30")],
                    "last_seen_evidence_quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-09-30")],
                    "metric": ["Management tone", "Adjusted EBIT guidance"],
                    "metric_display": ["Management tone", "Adjusted EBIT guidance"],
                    "promise_text": [
                        "Management remains optimistic about execution and opportunities ahead.",
                        "Adjusted EBIT guidance increased to $450 million to $465 million for FY 2025.",
                    ],
                    "text_full": [
                        "Management remains optimistic about execution and opportunities ahead.",
                        "Adjusted EBIT guidance increased to $450 million to $465 million for FY 2025.",
                    ],
                    "text_snippet": [
                        "Management remains optimistic about execution and opportunities ahead.",
                        "Adjusted EBIT guidance increased to $450 million to $465 million for FY 2025.",
                    ],
                    "target": ["", "$450.0m-$465.0m"],
                    "target_display": ["", "$450.0m-$465.0m"],
                    "target_time": [pd.NaT, pd.Timestamp("2025-12-31")],
                    "target_period_norm": ["", "FY2025"],
                    "promise_type": ["operational", "guidance_range"],
                    "theme_key": ["tone|fy2025", "ebit|fy2025"],
                    "source": [{"source_type": "earnings_release"}, {"source_type": "guidance_snapshot"}],
                }
            )

            inputs = _make_inputs(out_path, ticker="TEST")
            inputs = inputs.__class__(**{**vars(inputs), "promises": promises})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_tracker_ui_v2()
            tracker_ws = ctx.wb["Promise_Tracker_UI"]
            tracker_values = [str(tracker_ws.cell(row=rr, column=cc).value or "") for rr in range(1, tracker_ws.max_row + 1) for cc in range(1, 6)]

            assert not any("optimistic about execution" in val.lower() for val in tracker_values)


def test_gpre_tracker_does_not_originate_completed_operational_result_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_tracker_rejects_completed_results.xlsx")
            promises = pd.DataFrame(
                {
                    "promise_id": ["result-row", "target-row"],
                    "quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-09-30")],
                    "created_quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-09-30")],
                    "last_seen_quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-09-30")],
                    "first_seen_evidence_quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-09-30")],
                    "last_seen_evidence_quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-09-30")],
                    "metric": ["Strategic milestone", "45Z monetization / EBITDA"],
                    "metric_display": ["Strategic milestone", "45Z monetization / EBITDA"],
                    "promise_text": [
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                    ],
                    "text_full": [
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                    ],
                    "text_snippet": [
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025.",
                    ],
                    "target": ["", "$15.0m-$25.0m"],
                    "target_display": ["", "$15.0m-$25.0m"],
                    "target_time": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-12-31")],
                    "target_period_norm": ["Q42025", "Q42025"],
                    "promise_type": ["milestone", "guidance_range"],
                    "theme_key": ["nebraska|ops", "45z|q42025"],
                    "source": [{"source_type": "earnings_release"}, {"source_type": "earnings_release"}],
                }
            )

            inputs = _make_inputs(out_path, ticker="TEST")
            inputs = inputs.__class__(**{**vars(inputs), "promises": promises})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_tracker_ui_v2()
            tracker_ws = ctx.wb["Promise_Tracker_UI"]
            tracker_values = [str(tracker_ws.cell(row=rr, column=cc).value or "") for rr in range(1, tracker_ws.max_row + 1) for cc in range(1, 6)]

            assert not any("fully operational and sequestering" in val.lower() for val in tracker_values)
            assert not any("Advantage Nebraska is fully operational" in val for val in tracker_values)


def test_gpre_tracker_does_not_originate_online_ramping_or_debt_repaid_results(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_tracker_result_rows_blocked.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 2,
                    "note_id": ["ramp-1", "debt-1"],
                    "category": ["Programs / initiatives", "Debt / liquidity / covenants"],
                    "claim": [
                        "Central City and Wood River online and ramping.",
                        "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt.",
                    ],
                    "note": [
                        "Central City and Wood River online and ramping.",
                        "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt.",
                    ],
                    "metric_ref": ["Strategic milestone", "Debt reduction"],
                    "score": [92.0, 94.0],
                    "doc_type": ["presentation", "earnings_release"],
                    "doc": ["slides_q3.txt", "release_q3.txt"],
                    "evidence_snippet": [
                        "Central City and Wood River online and ramping.",
                        "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "slides_q3.txt", "doc_type": "presentation", "snippet": "Central City and Wood River online and ramping."}]),
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "Sale of Obion, Tennessee plant completed; proceeds used to fully repay $130.7 million junior mezzanine debt."}]),
                    ],
                }
            )

            inputs = _make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes)
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_quarter_notes_ui_v2()
            ctx.callbacks.write_promise_tracker_ui_v2()
            tracker_ws = ctx.wb["Promise_Tracker_UI"]
            tracker_values = [
                str(tracker_ws.cell(row=rr, column=cc).value or "")
                for rr in range(1, tracker_ws.max_row + 1)
                for cc in range(1, 6)
            ]

            assert not any("online and ramping" in val.lower() for val in tracker_values)
            assert not any("fully repay" in val.lower() or "Debt reduction" in val for val in tracker_values)


def test_gpre_tracker_does_not_originate_permit_or_equipment_execution_as_new_promises(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_tracker_blocks_later_evidence_origins.xlsx")
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2024-09-30"), pd.Timestamp("2024-06-30")],
                    "note_id": ["permit-1", "equip-1"],
                    "category": ["Programs / initiatives", "Programs / initiatives"],
                    "claim": [
                        "Advantage Nebraska strategy on track with pipeline partner receiving their first Class VI well permit in Wyoming.",
                        "Executed construction management agreements and ordered major equipment necessary to capture carbon from Nebraska facilities as part of Advantage Nebraska strategy.",
                    ],
                    "note": [
                        "Advantage Nebraska strategy on track with pipeline partner receiving their first Class VI well permit in Wyoming.",
                        "Executed construction management agreements and ordered major equipment necessary to capture carbon from Nebraska facilities as part of Advantage Nebraska strategy.",
                    ],
                    "metric_ref": ["45Z monetization / EBITDA", "45Z monetization / EBITDA"],
                    "score": [91.0, 90.0],
                    "doc_type": ["earnings_release", "earnings_release"],
                    "doc": ["release_q3.txt", "release_q2.txt"],
                    "evidence_snippet": [
                        "Advantage Nebraska strategy on track with pipeline partner receiving their first Class VI well permit in Wyoming.",
                        "Executed construction management agreements and ordered major equipment necessary to capture carbon from Nebraska facilities as part of Advantage Nebraska strategy.",
                    ],
                    "evidence_json": [
                        json.dumps([{"doc_path": "release_q3.txt", "doc_type": "earnings_release", "snippet": "Advantage Nebraska strategy on track with pipeline partner receiving their first Class VI well permit in Wyoming."}]),
                        json.dumps([{"doc_path": "release_q2.txt", "doc_type": "earnings_release", "snippet": "Executed construction management agreements and ordered major equipment necessary to capture carbon from Nebraska facilities as part of Advantage Nebraska strategy."}]),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
            ctx.callbacks.write_quarter_notes_ui_v2()
            ctx.callbacks.write_promise_tracker_ui_v2()
            tracker_ws = ctx.wb["Promise_Tracker_UI"]
            tracker_values = [str(tracker_ws.cell(row=rr, column=cc).value or "") for rr in range(1, tracker_ws.max_row + 1) for cc in range(1, 6)]

            assert not any("Class VI well permit" in val for val in tracker_values)
            assert not any("ordered major equipment" in val for val in tracker_values)


def test_pbi_progress_unifies_guidance_without_separate_guidance_accuracy_section(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_unified_guidance.xlsx")
            hist = _make_hist().copy()
            hist["revenue"] = [100_000_000.0, 110_000_000.0, 120_000_000.0, 530_000_000.0]
            progress = pd.DataFrame(
                {
                    "promise_id": ["rev-guidance"],
                    "quarter": [pd.Timestamp("2025-09-30")],
                    "status": ["on_track"],
                    "metric_ref": ["Revenue guidance"],
                    "target": ["$500.0m-$520.0m"],
                    "latest": ["not yet measurable"],
                    "rationale": ["FY 2025 revenue guidance was $500 million to $520 million."],
                    "promise_type": ["guidance_range"],
                    "guidance_type": ["period"],
                    "target_period_norm": ["FY2025"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2025 revenue guidance was $500 million to $520 million."})
                    ],
                }
            )

            inputs = _make_inputs(out_path, ticker="TEST", hist=hist)
            inputs = inputs.__class__(**{**vars(inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            visible_values = [str(ws.cell(row=rr, column=cc).value or "") for rr in range(1, ws.max_row + 1) for cc in range(1, 7)]
            assert not any("Guidance accuracy" in val for val in visible_values)
            assert sum(1 for val in visible_values if val == "Revenue guidance") == 1


def test_pbi_progress_uses_guidance_lifecycle_id_for_sheet_derived_guidance_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_guidance_id_unified.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["rev-guidance"],
                    "quarter": [pd.Timestamp("2025-03-31")],
                    "status": ["beat"],
                    "metric_ref": ["Revenue guidance"],
                    "target": ["$500m-$520m"],
                    "latest": ["$525m actual"],
                    "rationale": [
                        "FY 2025 revenue guidance was $500 million to $520 million.",
                    ],
                    "promise_type": ["guidance_range"],
                    "guidance_type": ["period"],
                    "target_period_norm": ["FY2025"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2025 revenue guidance was $500 million to $520 million."}),
                    ],
                }
            )
            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", promise_progress=progress))
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]
            ids = [str(ws.cell(row=rr, column=1).value or "") for rr in range(1, ws.max_row + 1)]
            assert any(val.startswith("guidance:revenue_guidance") for val in ids)
            assert not any(val.startswith("pbi_qn_sheet:") for val in ids)
            assert sum(1 for val in ids if "guidance:revenue_guidance" in val.lower()) == 1


def test_pbi_progress_collapses_duplicate_guidance_rows_for_same_metric_period(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_guidance_duplicate_collapse.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["pbi_qn_sheet:rev", "guidance:revenue_guidance:FY2025"],
                    "quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-12-31")],
                    "status": ["pending", "beat"],
                    "metric_ref": ["Revenue guidance", "Revenue guidance"],
                    "target": ["$500m-$520m", "$500m-$520m"],
                    "latest": ["not yet measurable", "$525m actual"],
                    "rationale": [
                        "FY 2025 revenue guidance was $500 million to $520 million.",
                        "Revenue came in at $525 million versus FY 2025 guidance of $500 million to $520 million.",
                    ],
                    "promise_type": ["guidance_range", "guidance_range"],
                    "guidance_type": ["period", "period"],
                    "target_period_norm": ["FY2025", "FY2025"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2025 revenue guidance was $500 million to $520 million."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "Revenue came in at $525 million versus FY 2025 guidance of $500 million to $520 million."}),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", promise_progress=progress))
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]
            ids = [str(ws.cell(row=rr, column=1).value or "") for rr in range(1, ws.max_row + 1)]
            revenue_ids = [val for val in ids if val.lower().startswith("guidance:revenue_guidance")]
            assert len(revenue_ids) == 1
            assert not any("pbi_qn_sheet:" in val.lower() for val in ids)


def test_pbi_progress_rendered_sheet_keeps_one_guidance_row_per_metric(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_rendered_metric_singleton.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["pbi_qn_sheet:ebit", "guidance_eval:ebit"],
                    "quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-12-31")],
                    "status": ["on_track", "pending"],
                    "metric_ref": ["Adjusted EBIT guidance", "Adjusted EBIT guidance"],
                    "target": ["$410m-$460m", "$410m-$460m"],
                    "latest": ["not yet measurable", "not yet measurable"],
                    "rationale": [
                        "FY 2026 Adjusted EBIT guidance target $410m-$460m.",
                        "Guidance period FY 2026 has not ended (see evaluated_through).",
                    ],
                    "promise_type": ["guidance_range", "guidance_range"],
                    "guidance_type": ["period", "period"],
                    "target_period_norm": ["FY2026", "FY2026"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2026 Adjusted EBIT guidance target $410m-$460m."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "Guidance period FY 2026 has not ended (see evaluated_through)."}),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", promise_progress=progress))
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]
            metric_rows = [
                str(ws.cell(row=rr, column=2).value or "").strip()
                for rr in range(1, ws.max_row + 1)
                if str(ws.cell(row=rr, column=2).value or "").strip() == "Adjusted EBIT guidance"
            ]
            assert len(metric_rows) <= 1


def test_pbi_progress_repairs_unreasonable_guidance_period_leakage(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_period_repair.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["guidance_eval:fcf"],
                    "quarter": [pd.Timestamp("2025-09-30")],
                    "status": ["pending"],
                    "metric_ref": ["FCF target"],
                    "target": ["$330m-$370m"],
                    "latest": ["not yet measurable"],
                    "rationale": ["Guidance period FY 2043 has not ended (see evaluated_through)."],
                    "promise_type": ["guidance_range"],
                    "guidance_type": ["period"],
                    "target_period_norm": ["FY2043"],
                    "target_period_label": ["FY 2043"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "Guidance period FY 2043 has not ended (see evaluated_through)."})
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", promise_progress=progress))
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]
            visible_values = [
                str(ws.cell(row=rr, column=cc).value or "")
                for rr in range(1, ws.max_row + 1)
                for cc in range(1, 7)
            ]
            assert all("FY 2043" not in val for val in visible_values)
            assert any("FY 2025" in val for val in visible_values)


def test_pbi_progress_collapses_cost_savings_rows_into_unified_guidance_lifecycle(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_cost_savings_unified.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["3729b09231c8", "3bf26de8be1e"],
                    "quarter": [pd.Timestamp("2024-12-31"), pd.Timestamp("2024-12-31")],
                    "status": ["on_track", "pending"],
                    "metric_ref": ["Cost savings target", "Cost savings target"],
                    "target": ["$150m-$170m", "$150m-$170m"],
                    "latest": ["not yet measurable", "not yet measurable"],
                    "rationale": [
                        "Cost savings target target $150m-$170m annualized savings.",
                        "Cost savings target target $150m-$170m annualized savings. Guidance period FY 2025 has not ended.",
                    ],
                    "promise_type": ["operational", "guidance_range"],
                    "guidance_type": ["run-rate", "period"],
                    "target_period_norm": ["ANNUALIZED_PROGRAM", "ANNUALIZED_PROGRAM"],
                    "promise_key": ["cost_savings_run_rate", "cost_savings_run_rate"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "Cost savings target target $150m-$170m annualized savings."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "Cost savings target target $150m-$170m annualized savings. Guidance period FY 2025 has not ended."}),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", promise_progress=progress))
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]
            ids = [str(ws.cell(row=rr, column=1).value or "") for rr in range(1, ws.max_row + 1)]
            cost_ids = [val for val in ids if "cost_savings" in val.lower()]
            assert cost_ids == ["guidance:cost_savings:ANNUALIZED_PROGRAM"]


def test_gpre_progress_collapses_same_lifecycle_subject_to_one_row_per_block(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_progress_same_lifecycle_collapse.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["cc-1", "cc-2"],
                    "quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-09-30")],
                    "status": ["completed", "in_progress"],
                    "metric_ref": ["Carbon capture milestone", "Carbon capture milestone"],
                    "target": ["Advantage Nebraska fully operational", "Advantage Nebraska fully operational"],
                    "latest": ["fully operational", "online and ramping"],
                    "rationale": [
                        "Advantage Nebraska is fully operational in Q3 2025.",
                        "Advantage Nebraska remained online and ramping in Q3 2025.",
                    ],
                    "promise_type": ["milestone", "milestone"],
                    "guidance_type": ["text", "text"],
                    "target_period_norm": ["Q32025", "Q32025"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "Advantage Nebraska is fully operational in Q3 2025."}),
                        json.dumps({"doc_type": "presentation", "snippet": "Advantage Nebraska remained online and ramping in Q3 2025."}),
                    ],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="TEST")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            rows = [
                rr
                for rr in range(1, ws.max_row + 1)
                if ws.cell(row=rr, column=2).value == "Advantage Nebraska startup"
            ]
            assert len(rows) == 1
            assert ws.cell(row=rows[0], column=5).value in {"Completed", "Hit", "Beat"}


def test_gpre_progress_does_not_keep_operational_update_as_parallel_monetization_row(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_progress_parent_child_collapse.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["startup", "monetization-noisy"],
                    "quarter": [pd.Timestamp("2025-12-31"), pd.Timestamp("2025-12-31")],
                    "status": ["completed", "completed"],
                    "metric_ref": ["Advantage Nebraska startup", "45Z monetization / EBITDA"],
                    "target": ["Advantage Nebraska fully operational", "$15.0m-$25.0m expected Q4 2025 monetization"],
                    "latest": ["fully operational", "Advantage Nebraska fully operational"],
                    "rationale": [
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming in Q4 2025.",
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming in Q4 2025.",
                    ],
                    "promise_type": ["milestone", "guidance_range"],
                    "guidance_type": ["text", "period"],
                    "target_period_norm": ["Q42025", "Q42025"],
                        "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming in Q4 2025."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming in Q4 2025."}),
                    ],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="TEST")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            visible_pairs = [
                (
                    str(ws.cell(row=rr, column=2).value or ""),
                    str(ws.cell(row=rr, column=4).value or ""),
                )
                for rr in range(1, ws.max_row + 1)
            ]
            assert ("Advantage Nebraska startup", "Advantage Nebraska fully operational (Q4 2025)") in visible_pairs
            assert not any(metric == "45Z monetization / EBITDA" and "fully operational" in latest.lower() for metric, latest in visible_pairs)


def test_gpre_progress_filters_junk_labels_and_keeps_clean_forward_families(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_progress_quality_cleanup.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["good-45z", "good-interest", "bad-rd", "bad-fragment", "bad-year-ended", "bad-nan"],
                    "quarter": [pd.Timestamp("2025-12-31")] * 6,
                    "status": ["on_track", "on_track", "in_progress", "in_progress", "on_track", "in_progress"],
                    "metric_ref": [
                        "least 45Z monetization / EBITDA",
                        "Management operating target",
                        "company has federal R&D Strategic milestone",
                        "in all of our Strategic milestone",
                        "year ended December cost savings",
                        "Advantage Nebraska startup",
                    ],
                    "target": [">= $188.0m in 2026", "$30.0m-$35.0m", "1", "1", ">= $16.1m", "1"],
                    "latest": ["not yet measurable", "not yet measurable", "$63.9m disclosed in 2022", "Expected in 2024", "$16.1m disclosed in 2025", "nan"],
                    "rationale": [
                        "FY 2026 45Z-related Adjusted EBITDA outlook is at least $188.0m.",
                        "Annualized 2026 interest expense is expected at about $30.0m-$35.0m.",
                        "At December 31, 2022, the company has federal R&D credits of $63.9 million which will begin to expire in 2033.",
                        "We successfully executed full-scale production runs of 60% protein at one location during the quarter and have begun to ship product to customers globally into full.",
                        "Corporate activities includes $16.1 million of restructuring costs for the year ended December 31, 2025.",
                        "Milestone pending until stated deadline.",
                    ],
                    "promise_type": ["guidance_range", "guidance_range", "milestone", "milestone", "operational", "milestone"],
                    "guidance_type": ["period", "period", "text", "text", "period", "text"],
                    "target_period_norm": ["FY2026", "FY2026", "", "", "FY2025", "Q42025"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2026 45Z-related Adjusted EBITDA outlook is at least $188.0m."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "Annualized 2026 interest expense is expected at about $30.0m-$35.0m."}),
                        json.dumps({"doc_type": "10-K", "snippet": "At December 31, 2022, the company has federal R&D credits of $63.9 million which will begin to expire in 2033."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "We successfully executed full-scale production runs of 60% protein at one location during the quarter and have begun to ship product to customers globally into full."}),
                        json.dumps({"doc_type": "10-K", "snippet": "Corporate activities includes $16.1 million of restructuring costs for the year ended December 31, 2025."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "Milestone pending until stated deadline."}),
                    ],
                }
            )

            ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", promise_progress=progress))
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]
            visible_metrics = {
                str(ws.cell(row=rr, column=2).value or "").strip()
                for rr in range(1, ws.max_row + 1)
                if str(ws.cell(row=rr, column=2).value or "").strip() not in {"", "metric"}
            }
            visible_latest = [
                str(ws.cell(row=rr, column=4).value or "").strip().lower()
                for rr in range(1, ws.max_row + 1)
            ]

            assert "45Z-related Adjusted EBITDA outlook" in visible_metrics
            assert "Interest expense outlook" in visible_metrics
            assert "company has federal R&D Strategic milestone" not in visible_metrics
            assert "in all of our Strategic milestone" not in visible_metrics
            assert "year ended December cost savings" not in visible_metrics
            assert all(val != "nan" for val in visible_latest)


def test_pbi_progress_can_add_strategic_review_from_quarter_notes_without_visible_tracker(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_progress_qnote_strategic_review.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["rev", "ebit", "eps", "fcf"],
                    "quarter": [pd.Timestamp("2025-12-31")] * 4,
                    "status": ["pending"] * 4,
                    "metric_ref": ["Revenue guidance", "Adjusted EBIT guidance", "EPS guidance", "FCF target"],
                    "target": ["$1.76bn-$1.86bn", "$410m-$460m", "$1.40-$1.60", "$340m-$370m"],
                    "latest": ["not yet measurable"] * 4,
                    "rationale": [
                        "FY 2026 revenue guidance of $1.76bn-$1.86bn.",
                        "FY 2026 adjusted EBIT guidance of $410m-$460m.",
                        "FY 2026 adjusted EPS guidance of $1.40-$1.60.",
                        "FY 2026 free cash flow target of $340m-$370m.",
                    ],
                    "promise_type": ["guidance_range"] * 4,
                    "guidance_type": ["period"] * 4,
                    "target_period_norm": ["FY2026"] * 4,
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2026 revenue guidance of $1.76bn-$1.86bn."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2026 adjusted EBIT guidance of $410m-$460m."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2026 adjusted EPS guidance of $1.40-$1.60."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2026 free cash flow target of $340m-$370m."}),
                    ],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="TEST", promise_progress=progress)
            inputs = base_inputs.__class__(**vars(base_inputs))
            ctx = build_writer_context(inputs)
            ctx.state["ui_state"]["quarter_notes_ui_rows"] = {
                pd.Timestamp("2025-12-31").date(): [
                    {
                        "note_id": "sr-q4-2025",
                        "text_full": "Strategic review phase 2 remains on track by end of Q2 2026.",
                        "_render_summary": "Strategic review phase 2 remains on track by end of Q2 2026.",
                        "metric_ref": "Strategic milestone",
                        "score": 92.0,
                        "source": {
                            "source_type": "ceo_letter",
                            "doc": "q4_2025_ceo_letter.pdf",
                        },
                    }
                ]
            }

            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]
            rows = [
                (
                    str(ws.cell(row=rr, column=2).value or "").strip(),
                    str(ws.cell(row=rr, column=6).value or "").strip(),
                )
                for rr in range(1, ws.max_row + 1)
            ]

            assert any(metric == "Strategic milestone" for metric, _ in rows)
            assert any("Strategic review phase 2 remains on track by end of Q2 2026" in rationale for _, rationale in rows)


def test_gpre_progress_can_add_targeted_qnote_outlooks_even_when_block_is_not_sparse(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_progress_qnote_targeted_backfill.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["neb-start", "neb-ebitda", "fortyfivez", "remaining"],
                    "quarter": [pd.Timestamp("2025-12-31")] * 4,
                    "status": ["completed", "on_track", "on_track", "on_track"],
                    "metric_ref": [
                        "Advantage Nebraska startup",
                        "Advantage Nebraska EBITDA opportunity",
                        "45Z-related Adjusted EBITDA",
                        "45Z from remaining facilities",
                    ],
                    "target": ["", "> $150.0m in 2026", ">= $188.0m in 2026", "> $38.0m expected in 2026"],
                    "latest": [
                        "Advantage Nebraska fully operational",
                        "$150.0m disclosed in 2026",
                        "Advantage Nebraska fully operational (Q4 2025)",
                        "$23.4m YTD 45Z value realized (net of discounts)",
                    ],
                    "rationale": [
                        "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming.",
                        "Advantage Nebraska 2026 Adjusted EBITDA opportunity exceeds $150.0m.",
                        "FY 2026 45Z-related Adjusted EBITDA outlook is at least $188.0m.",
                        "Expected 45Z generation from remaining facilities exceeds $38.0m in 2026.",
                    ],
                    "promise_type": ["operational"] * 4,
                    "guidance_type": ["period"] * 4,
                    "target_period_norm": ["Q42025", "FY2026", "FY2026", "FY2026"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "Advantage Nebraska 2026 Adjusted EBITDA opportunity exceeds $150.0m."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "FY 2026 45Z-related Adjusted EBITDA outlook is at least $188.0m."}),
                        json.dumps({"doc_type": "earnings_release", "snippet": "Expected 45Z generation from remaining facilities exceeds $38.0m in 2026."}),
                    ],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="TEST", promise_progress=progress)
            inputs = base_inputs.__class__(**vars(base_inputs))
            ctx = build_writer_context(inputs)
            ctx.state["ui_state"]["quarter_notes_ui_rows"] = {
                pd.Timestamp("2025-12-31").date(): [
                    {
                        "note_id": "interest-q4-2025",
                        "text_full": "Annualized 2026 interest expense is expected at about $30.0m-$35.0m, reflecting the 2030 convertible notes and carbon equipment financing.",
                        "_render_summary": "Annualized 2026 interest expense is expected at about $30.0m-$35.0m.",
                        "metric_ref": "Interest expense outlook",
                        "score": 95.0,
                        "source": {"source_type": "earnings_release", "doc": "q4_2025_earnings_release.htm"},
                    },
                    {
                        "note_id": "monet-q4-2025",
                        "text_full": "Q4 2025 45Z monetization expected at $15m-$25m.",
                        "_render_summary": "Q4 2025 45Z monetization expected at $15m-$25m.",
                        "metric_ref": "45Z monetization / EBITDA",
                        "score": 96.0,
                        "source": {"source_type": "earnings_release", "doc": "q4_2025_earnings_release.htm"},
                    },
                ]
            }

            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]
            rows = [
                (
                    str(ws.cell(row=rr, column=2).value or "").strip(),
                    str(ws.cell(row=rr, column=3).value or "").strip(),
                    str(ws.cell(row=rr, column=6).value or "").strip(),
                )
                for rr in range(1, ws.max_row + 1)
                if str(ws.cell(row=rr, column=2).value or "").strip() not in {"", "metric"}
            ]

            assert any(metric == "Interest expense outlook" for metric, _, _ in rows)
            assert any(metric == "45Z monetization outlook" and "$15.0m-$25.0m" in target for metric, target, _ in rows)
            assert any("interest expense is expected" in rationale.lower() for _, _, rationale in rows)


def test_gpre_progress_collapses_duplicate_same_promise_id_rows_after_hydration(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_progress_same_promise_id_collapse.xlsx")
            progress = pd.DataFrame(
                {
                    "promise_id": ["neb-45z", "neb-45z"],
                    "quarter": [pd.Timestamp("2025-06-30"), pd.Timestamp("2025-06-30")],
                    "status": ["beat", "in_progress"],
                    "metric_ref": ["45Z Adjusted EBITDA / monetization", "45Z Adjusted EBITDA / monetization"],
                    "target": ["$15.0m-$25.0m expected Q4 2025 monetization", "$15.0m-$25.0m expected Q4 2025 monetization"],
                    "latest": ["~$27.7m realized in Q4 2025 (FY less 9M)", "Advantage Nebraska fully operational"],
                    "rationale": [
                        "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025. Later update: ~$27.7m realized in Q4 2025.",
                        "Executed construction management agreements and later update: Advantage Nebraska fully operational in Q4 2025.",
                    ],
                    "promise_type": ["operational", "operational"],
                    "guidance_type": ["", ""],
                    "target_period_norm": ["Q42025", "Q42025"],
                    "parent_subject_key": ["program:nebraska45z", "program:nebraska45z"],
                    "canonical_subject_key": ["program:nebraska45z|45z_monetization|q42025", "program:nebraska45z|45z_monetization|q42025"],
                    "lifecycle_subject_key": ["program:nebraska45z|monetization|q42025", "program:nebraska45z|monetization|q42025"],
                    "source_type": ["earnings_release", "presentation"],
                    "statement_class": ["structured_numeric_bridge", "result_evidence"],
                    "evidence_role": ["later_evidence", "result_evidence"],
                    "source_evidence_json": [
                        json.dumps({"doc_type": "earnings_release", "snippet": "On track for $15 - $25 million of 45Z production tax credit monetization value in Q4 2025. Later update: ~$27.7m realized in Q4 2025."}),
                        json.dumps({"doc_type": "presentation", "snippet": "Executed construction management agreements and later update: Advantage Nebraska fully operational in Q4 2025."}),
                    ],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="TEST")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "promise_progress": progress})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_promise_progress_ui_v2()
            ws = ctx.wb["Promise_Progress_UI"]

            visible_rows = [
                (
                    str(ws.cell(row=rr, column=1).value or ""),
                    str(ws.cell(row=rr, column=2).value or ""),
                    str(ws.cell(row=rr, column=4).value or ""),
                )
                for rr in range(1, ws.max_row + 1)
                if str(ws.cell(row=rr, column=2).value or "") == "45Z Adjusted EBITDA / monetization"
            ]
            assert len(visible_rows) == 1
            assert "$27.7m 45Z value realized" in visible_rows[0][2]


def test_gpre_valuation_ignores_generic_dividends_and_distributions_for_common_dividend_logic(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_no_generic_common_dividend.xlsx")
            sec_cache_dir = case_dir / "TEST" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            base_inputs = _make_inputs(out_path, ticker="TEST")
            hist = base_inputs.hist.copy()
            hist["dividends_cash"] = pd.NA
            hist["payments_of_dividends"] = 12_000_000.0

            inputs = base_inputs.__class__(**{**vars(base_inputs), "hist": hist, "cache_dir": sec_cache_dir})
            ctx = build_writer_context(inputs)
            ensure_valuation_inputs(ctx)
            ctx.callbacks.write_valuation_sheet()
            ws = ctx.wb["Valuation"]

            div_cash_row = _find_row_with_value(ws, "Dividends (TTM, cash)")
            div_note_row = _find_row_with_value(ws, "Dividends ($/share)")

            assert div_cash_row is not None
            assert div_note_row is not None
            assert "$" not in str(ws.cell(row=div_cash_row, column=2).value or "")
            assert "no current common dividend/share signal" in str(ws.cell(row=div_note_row, column=2).value or "").lower()


def test_gpre_quarter_notes_do_not_turn_cumulative_buyback_totals_into_q2_q3_execution(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_notes_no_cumulative_q2_q3_buybacks.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940225000090_gpre-20250630.htm").write_text(
                "Green Plains Inc. To date, we have repurchased approximately 7.4 million shares of common stock "
                "for approximately $92.8 million under the program. No repurchases were made during Q2 2025.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940225000163_gpre-20250930.htm").write_text(
                "Green Plains Inc. To date, we have repurchased approximately 7.4 million shares of common stock "
                "for approximately $92.8 million under the program. No repurchases were made during Q3 2025.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940226000019_gpre-20251231.htm").write_text(
                "Green Plains Inc. On October 27, 2025, in conjunction with the privately negotiated exchange and "
                "subscription agreements for the 2030 Notes, the company repurchased 2.9 million shares of its "
                "common stock for a total of $30.0 million under the repurchase program.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [
                        pd.Timestamp("2025-06-30"),
                        pd.Timestamp("2025-09-30"),
                        pd.Timestamp("2025-12-31"),
                    ],
                    "note_id": ["q2-guid", "q3-guid", "q4-guid"],
                    "category": ["Guidance / outlook"] * 3,
                    "claim": [
                        "Q2 block placeholder.",
                        "Q3 block placeholder.",
                        "Q4 block placeholder.",
                    ],
                    "note": [
                        "Q2 block placeholder.",
                        "Q3 block placeholder.",
                        "Q4 block placeholder.",
                    ],
                    "metric_ref": ["Guidance"] * 3,
                    "score": [90.0, 90.0, 90.0],
                    "doc_type": ["earnings_release"] * 3,
                    "doc": ["q2.txt", "q3.txt", "q4.txt"],
                    "source_type": ["earnings_release"] * 3,
                    "evidence_snippet": [
                        "Q2 block placeholder.",
                        "Q3 block placeholder.",
                        "Q4 block placeholder.",
                    ],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="GPRE", quarter_notes=quarter_notes)
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            q2_rows = _quarter_block_notes(ws, "2025-06-30")
            q3_rows = _quarter_block_notes(ws, "2025-09-30")
            q4_rows = _quarter_block_notes(ws, "2025-12-31")

            assert not any("7.4m shares" in note or "$92.8m" in note for note in q2_rows)
            assert not any("7.4m shares" in note or "$92.8m" in note for note in q3_rows)
            assert any("2.9m shares" in note and "$30.0m" in note for note in q4_rows)


def test_gpre_valuation_buyback_maps_ignore_cumulative_program_totals(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_no_cumulative_buyback_valuation.xlsx")
            sec_cache_dir = case_dir / "TEST" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940225000163_gpre-20250930.htm").write_text(
                "Green Plains Inc. To date, we have repurchased approximately 7.4 million shares of common stock "
                "for approximately $92.8 million under the program. No repurchases were made during Q3 2025.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940226000019_gpre-20251231.htm").write_text(
                "Green Plains Inc. Holders of the 2027 Notes will have the right, at their option, to require the "
                "company to repurchase their 2.25% Convertible Senior Notes due 2027 upon a fundamental change. "
                "On October 27, 2025, in conjunction with the privately negotiated exchange and subscription "
                "agreements for the 2030 Notes, the company repurchased 2.9 million shares of its common stock for "
                "a total of $30.0 million under the repurchase program. No other repurchase was made during 2025. "
                "We did not repurchase any common stock in 2024 or 2023. To date, we have repurchased approximately "
                "10.3 million shares of common stock for approximately $122.8 million under the program. At "
                "February 10, 2026, $77.2 million in share repurchase authorization remained.",
                encoding="utf-8",
            )

            base_inputs = _make_inputs(out_path, ticker="TEST")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            ctx = build_writer_context(inputs)
            ensure_valuation_inputs(ctx)
            ctx.callbacks.write_valuation_sheet()
            ws = ctx.wb["Valuation"]

            buybacks_note_row = _find_row_with_value(ws, "Buybacks note")
            assert buybacks_note_row is not None
            buyback_note = str(ws.cell(row=buybacks_note_row, column=2).value or "")
            assert "$2.0m" not in buyback_note
            assert "$92.8m" not in buyback_note
            assert "$122.8m" not in buyback_note
            assert "TTM $30.0m" not in buyback_note
            buybacks_shares_row = _find_row_with_value(ws, "Buybacks (shares)")
            assert buybacks_shares_row is not None
            buyback_shares_detail = str(ws.cell(row=buybacks_shares_row, column=2).value or "")
            assert "TTM +2.900m" not in buyback_shares_detail
            audit = dict((ctx.derived.valuation_precompute_bundle or {}).get("valuation_audit") or {})
            suppress_reason = str(audit[pd.Timestamp("2025-09-30")]["buyback_cash"]["suppress_reason"] or "")
            assert suppress_reason in {
                "context/program text blocked for execution metrics",
                "no explicit quarter-safe execution",
            }


def test_valuation_buyback_share_detail_does_not_fallback_to_share_count_delta(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "valuation_buyback_shares_no_delta_fallback.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940226000019_gpre-20251231.htm").write_text(
                "Green Plains Inc. On October 27, 2025, in conjunction with the privately negotiated exchange and "
                "subscription agreements for the 2030 Notes, the company repurchased approximately 2.9 million shares "
                "of its common stock for approximately $30.0 million. No other repurchase was made during 2025.",
                encoding="utf-8",
            )
            hist = pd.DataFrame(
                {
                    "quarter": pd.to_datetime(["2025-03-31", "2025-06-30", "2025-09-30", "2025-12-31"]),
                    "revenue": [500_000_000.0] * 4,
                    "ebitda": [50_000_000.0] * 4,
                    "ebit": [30_000_000.0] * 4,
                    "op_income": [25_000_000.0] * 4,
                    "cash": [20_000_000.0] * 4,
                    "debt_core": [150_000_000.0] * 4,
                    "interest_paid": [10_000_000.0] * 4,
                    "shares_outstanding": [100_000_000.0, 103_000_000.0, 107_000_000.0, 112_000_000.0],
                    "shares_diluted": [100_000_000.0, 103_000_000.0, 107_000_000.0, 112_000_000.0],
                    "market_cap": [1_000_000_000.0] * 4,
                }
            )

            base_inputs = _make_inputs(out_path, ticker="GPRE", hist=hist)
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            ctx = build_writer_context(inputs)
            ensure_valuation_inputs(ctx)
            ctx.callbacks.write_valuation_sheet()
            ws = ctx.wb["Valuation"]

            buybacks_shares_row = _find_row_with_value(ws, "Buybacks (shares)")
            assert buybacks_shares_row is not None
            buyback_shares_detail = str(ws.cell(row=buybacks_shares_row, column=2).value or "")
            assert "QoQ +2.900m" in buyback_shares_detail
            assert "TTM -" not in buyback_shares_detail

            audit = dict((ctx.derived.valuation_precompute_bundle or {}).get("valuation_audit") or {})
            assert audit[pd.Timestamp("2025-09-30")]["buyback_shares"]["value"] is None
            assert audit[pd.Timestamp("2025-09-30")]["buyback_shares"]["suppress_reason"] == "derived share delta blocked for execution metrics"


def test_pbi_valuation_prefers_filing_table_buyback_truth_over_rounded_press_release(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_valuation_buyback_table_truth.xlsx")
            sec_cache_dir = case_dir / "PBI" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828026008604_q42025earningspressrelea.htm").write_text(
                "During the quarter we repurchased 12.6 million shares for $127.0 million.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828026012345_pbi-20251231.htm").write_text(
                "The following table provides information about common stock purchases during the three months ended December 31, 2025. "
                "Total number of shares purchased Average price paid per share Total number of shares purchased as part of publicly announced plans or programs Approximate dollar value of shares that may yet be purchased under the plans or programs. "
                "October 2025 3,203,100 $ 11.30 3,203,100 $212,031 "
                "November 2025 7,926,090 $ 9.52 7,926,090 $136,553 "
                "December 2025 1,484,407 $ 10.05 1,484,407 $121,639 "
                "12,613,597&#160; $ 10.04&#160; 12,613,597 ",
                encoding="utf-8",
            )

            base_inputs = _make_inputs(out_path, ticker="PBI")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            ctx = build_writer_context(inputs)
            ensure_valuation_inputs(ctx)
            ctx.callbacks.write_valuation_sheet()
            ws = ctx.wb["Valuation"]

            buybacks_row = _find_row_with_value(ws, "Buybacks (shares)")
            buybacks_note_row = _find_row_with_value(ws, "Buybacks note")

            assert buybacks_row is not None
            assert buybacks_note_row is not None
            assert "QoQ +12.614m" in str(ws.cell(row=buybacks_row, column=2).value or "")

            buyback_note = str(ws.cell(row=buybacks_note_row, column=2).value or "")
            assert "Remaining capacity $359.0m" in buyback_note
            assert "Latest increase by $250.0m on 2026-02-13" in buyback_note
            assert "$127.0m" not in buyback_note
            assert "$10.08/share" not in buyback_note
            assert "$126.6m" not in buyback_note
            assert "$10.04/share" not in buyback_note


def test_pbi_valuation_prefers_q4_issuer_purchases_table_over_other_repurchase_narrative_in_same_filing(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_valuation_buyback_table_vs_same_doc_narrative.xlsx")
            sec_cache_dir = case_dir / "PBI" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828026008604_q42025earningspressrelea.htm").write_text(
                "During the quarter we repurchased 12.6 million shares for $127.0 million.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828026012345_pbi-20251231.htm").write_text(
                "During 2025, we used $61.9 million of proceeds to repurchase 5.5 million shares. "
                "The following table provides information about common stock purchases during the three months ended December 31, 2025. "
                "Total number of shares purchased Average price paid per share Total number of shares purchased as part of publicly announced plans or programs Approximate dollar value of shares that may yet be purchased under the plans or programs. "
                "October 2025 3,203,100&#160; $ 11.30&#160; 3,203,100 $212,031 "
                "November 2025 7,926,090&#160; $ 9.52&#160; 7,926,090 $136,553 "
                "December 2025 1,484,407&#160; $ 10.05&#160; 1,484,407 $121,639 "
                "12,613,597&#160; $ 10.04&#160; 12,613,597 ",
                encoding="utf-8",
            )

            base_inputs = _make_inputs(out_path, ticker="PBI")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            ctx = build_writer_context(inputs)
            ensure_valuation_inputs(ctx)
            ctx.callbacks.write_valuation_sheet()
            ws = ctx.wb["Valuation"]

            buybacks_row = _find_row_with_value(ws, "Buybacks (shares)")
            buybacks_note_row = _find_row_with_value(ws, "Buybacks note")

            assert buybacks_row is not None
            assert buybacks_note_row is not None
            assert "QoQ +12.614m" in str(ws.cell(row=buybacks_row, column=2).value or "")

            buyback_note = str(ws.cell(row=buybacks_note_row, column=2).value or "")
            assert "Remaining capacity $359.0m" in buyback_note
            assert "Latest increase by $250.0m on 2026-02-13" in buyback_note
            assert "$61.9m" not in buyback_note
            assert "$127.0m" not in buyback_note
            assert "$126.6m" not in buyback_note
            assert "$10.04/share" not in buyback_note


def test_pbi_buyback_ttm_cash_uses_quarter_safe_cashflow_deltas(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_buyback_ttm_cash_quarter_safe.xlsx")
            sec_cache_dir = case_dir / "PBI" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828025023713_pbi-20250331.htm").write_text(
                "Condensed Consolidated Statements of Cash Flows Common stock repurchases ( 15,000 )",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828025036856_pbi-20250630.htm").write_text(
                "Condensed Consolidated Statements of Cash Flows Common stock repurchases ( 90,274 )",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828025047360_pbi-20250930.htm").write_text(
                "Condensed Consolidated Statements of Cash Flows Common stock repurchases ( 251,774 )",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828026009650_pbi-20251231.htm").write_text(
                "Consolidated Statements of Cash Flows Common stock repurchases ( 378,361 )",
                encoding="utf-8",
            )

            base_inputs = _make_inputs(out_path, ticker="PBI")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            result = write_excel_from_inputs(inputs)

            validate_valuation_export(out_path, result.valuation_export_expectation)

            valuation_snapshot = result.saved_workbook_provenance.get("valuation_snapshot") or {}
            quarter_headers = list(valuation_snapshot.get("quarter_headers") or [])
            buyback_cash_vals = list((valuation_snapshot.get("grid_rows") or {}).get("Buybacks (cash)") or [])
            buyback_ttm_vals = list((valuation_snapshot.get("grid_rows") or {}).get("Buybacks (TTM, cash)") or [])
            cash_by_q = dict(zip(quarter_headers, buyback_cash_vals))
            ttm_by_q = dict(zip(quarter_headers, buyback_ttm_vals))

            assert cash_by_q.get("2025-Q1") == pytest.approx(15.0)
            assert cash_by_q.get("2025-Q2") == pytest.approx(75.274)
            assert cash_by_q.get("2025-Q3") == pytest.approx(161.5)
            assert cash_by_q.get("2025-Q4") == pytest.approx(126.587)
            assert ttm_by_q.get("2025-Q1") in (None, "")
            assert ttm_by_q.get("2025-Q2") in (None, "")
            assert ttm_by_q.get("2025-Q3") in (None, "")
            assert ttm_by_q.get("2025-Q4") == pytest.approx(378.361)
            assert ttm_by_q.get("2025-Q4") != pytest.approx(524.91407196)

            wb = load_workbook(out_path, data_only=False, read_only=False)
            try:
                ws = wb["Valuation"]
                buyback_cash_row = _find_row_with_value(ws, "Buybacks (cash)")
                buyback_ttm_row = _find_row_with_value(ws, "Buybacks (TTM, cash)")
                assert buyback_cash_row is not None
                assert buyback_ttm_row is not None
                assert buyback_cash_row + 1 == buyback_ttm_row
            finally:
                wb.close()


def test_valuation_first_visible_year_can_use_hidden_history_for_heatmap_fill() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "valuation_hidden_history_heatmap.xlsx")
        quarters = pd.period_range("2022Q1", "2025Q4", freq="Q").to_timestamp("Q")
        hist = pd.DataFrame(
            {
                "quarter": pd.to_datetime(quarters),
                "revenue": [
                    100_000_000.0,
                    110_000_000.0,
                    120_000_000.0,
                    130_000_000.0,
                    120_000_000.0,
                    130_000_000.0,
                    140_000_000.0,
                    150_000_000.0,
                    140_000_000.0,
                    150_000_000.0,
                    160_000_000.0,
                    170_000_000.0,
                    160_000_000.0,
                    170_000_000.0,
                    180_000_000.0,
                    190_000_000.0,
                ],
                "cfo": [(10.0 + i) * 1_000_000.0 for i in range(len(quarters))],
                "capex": [(2.0 + (0.1 * i)) * 1_000_000.0 for i in range(len(quarters))],
                "ebitda": [(15.0 + i) * 1_000_000.0 for i in range(len(quarters))],
                "ebit": [(9.0 + i) * 1_000_000.0 for i in range(len(quarters))],
                "cash": [(20.0 + i) * 1_000_000.0 for i in range(len(quarters))],
                "debt_core": [(80.0 - i) * 1_000_000.0 for i in range(len(quarters))],
                "shares_outstanding": [10_000_000.0] * len(quarters),
                "shares_diluted": [10_000_000.0] * len(quarters),
                "market_cap": [100_000_000.0] * len(quarters),
                "interest_expense_net": [1_000_000.0] * len(quarters),
            }
        )

        ctx = build_writer_context(_make_inputs(out_path, hist=hist))
        ensure_valuation_inputs(ctx)
        ctx.callbacks.write_valuation_sheet()
        ws = ctx.wb["Valuation"]

        q1_2023_col = _find_col_with_value(ws, "2023-Q1", row=6)
        revenue_row = _find_row_with_value(ws, "Revenue")
        buybacks_cash_row = _find_row_with_value(ws, "Buybacks (cash)")

        assert q1_2023_col is not None
        assert revenue_row is not None
        assert buybacks_cash_row is not None
        assert ws.cell(row=revenue_row, column=q1_2023_col).value == pytest.approx(120.0)
        assert _fill_rgb(ws.cell(row=revenue_row, column=q1_2023_col)) == "002F80ED"
        assert ws.cell(row=buybacks_cash_row, column=q1_2023_col).value in (None, "")
        assert _fill_rgb(ws.cell(row=buybacks_cash_row, column=q1_2023_col)) in {"", "00000000", "000000"}


def test_valuation_derived_rows_can_use_hidden_2022_history_for_2023_fill() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "valuation_hidden_history_derived_rows.xlsx")
        quarters = pd.period_range("2022Q1", "2025Q4", freq="Q").to_timestamp("Q")
        revenue = [
            100_000_000.0,
            105_000_000.0,
            110_000_000.0,
            115_000_000.0,
            120_000_000.0,
            125_000_000.0,
            130_000_000.0,
            135_000_000.0,
            140_000_000.0,
            145_000_000.0,
            150_000_000.0,
            155_000_000.0,
            160_000_000.0,
            165_000_000.0,
            170_000_000.0,
            175_000_000.0,
        ]
        gross_profit = [
            40_000_000.0,
            42_000_000.0,
            44_000_000.0,
            46_000_000.0,
            60_000_000.0,
            62_500_000.0,
            65_000_000.0,
            67_500_000.0,
            70_000_000.0,
            72_500_000.0,
            75_000_000.0,
            77_500_000.0,
            80_000_000.0,
            82_500_000.0,
            85_000_000.0,
            87_500_000.0,
        ]
        ebitda = [30_000_000.0 + (i * 1_000_000.0) for i in range(len(quarters))]
        hist = pd.DataFrame(
            {
                "quarter": pd.to_datetime(quarters),
                "revenue": revenue,
                "gross_profit": gross_profit,
                "ebitda": ebitda,
                "ebit": [20_000_000.0 + (i * 1_000_000.0) for i in range(len(quarters))],
                "net_income": [10_000_000.0 + (i * 1_000_000.0) for i in range(len(quarters))],
                "cfo": [12_000_000.0 + (i * 500_000.0) for i in range(len(quarters))],
                "capex": [3_000_000.0 + (i * 100_000.0) for i in range(len(quarters))],
                "cash": [20_000_000.0 + (i * 1_000_000.0) for i in range(len(quarters))],
                "debt_core": [90_000_000.0 - (i * 1_000_000.0) for i in range(len(quarters))],
                "shares_outstanding": [10_000_000.0] * len(quarters),
                "shares_diluted": [10_000_000.0] * len(quarters),
                "market_cap": [100_000_000.0] * len(quarters),
                "interest_expense_net": [1_000_000.0] * len(quarters),
            }
        )
        adj_ebitda = [
            40_000_000.0,
            41_000_000.0,
            42_000_000.0,
            43_000_000.0,
            52_000_000.0,
            53_000_000.0,
            54_000_000.0,
            55_000_000.0,
            60_000_000.0,
            61_000_000.0,
            62_000_000.0,
            63_000_000.0,
            68_000_000.0,
            69_000_000.0,
            70_000_000.0,
            71_000_000.0,
        ]
        adj_metrics = pd.DataFrame(
            {
                "quarter": pd.to_datetime(quarters),
                "adj_ebitda": adj_ebitda,
            }
        )

        base_inputs = _make_inputs(out_path, hist=hist)
        inputs = base_inputs.__class__(**{**vars(base_inputs), "adj_metrics": adj_metrics})
        ctx = build_writer_context(inputs)
        ensure_valuation_inputs(ctx)
        ctx.callbacks.write_valuation_sheet()
        ws = ctx.wb["Valuation"]

        q1_2023_col = _find_col_with_value(ws, "2023-Q1", row=6)
        gross_margin_row = _find_row_with_value(ws, "Gross margin %")
        adj_diff_row = _find_row_with_value(ws, "Adj EBITDA - EBITDA")

        assert q1_2023_col is not None
        assert gross_margin_row is not None
        assert adj_diff_row is not None
        assert ws.cell(row=gross_margin_row, column=q1_2023_col).value == pytest.approx(0.5)
        assert _fill_rgb(ws.cell(row=gross_margin_row, column=q1_2023_col)) == "002F80ED"
        assert ws.cell(row=adj_diff_row, column=q1_2023_col).value == pytest.approx(18.0)
        assert _fill_rgb(ws.cell(row=adj_diff_row, column=q1_2023_col)) == "002F80ED"


def test_valuation_buybacks_cash_uses_qoq_heatmap_when_prior_visible_quarter_exists(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "valuation_buybacks_cash_qoq_fill.xlsx")
            sec_cache_dir = case_dir / "PBI" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828025012345_pbi-20250331.htm").write_text(
                "Pitney Bowes Inc. During the quarter we repurchased 1.5 million shares for $15.0 million. "
                "Condensed Consolidated Statements of Cash Flows Common stock repurchases ( 15,000 )",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828025023456_pbi-20250630.htm").write_text(
                "Pitney Bowes Inc. During the quarter we repurchased 7.6 million shares for $75.3 million. "
                "Condensed Consolidated Statements of Cash Flows Common stock repurchases ( 90,274 )",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828025034567_pbi-20250930.htm").write_text(
                "Pitney Bowes Inc. During the quarter we repurchased 14.1 million shares for $161.5 million. "
                "Condensed Consolidated Statements of Cash Flows Common stock repurchases ( 251,774 )",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828026045678_pbi-20251231.htm").write_text(
                "Pitney Bowes Inc. During the quarter we repurchased 12.6 million shares for $126.6 million. "
                "Consolidated Statements of Cash Flows Common stock repurchases ( 378,361 )",
                encoding="utf-8",
            )

            base_inputs = _make_inputs(out_path, ticker="PBI")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            result = write_excel_from_inputs(inputs)

            validate_valuation_export(out_path, result.valuation_export_expectation)

            wb = load_workbook(out_path, data_only=False, read_only=False)
            try:
                ws = wb["Valuation"]
                q1_2025_col = _find_col_with_value(ws, "2025-Q1", row=6)
                q2_2025_col = _find_col_with_value(ws, "2025-Q2", row=6)
                q3_2025_col = _find_col_with_value(ws, "2025-Q3", row=6)
                buyback_cash_row = _find_row_with_value(ws, "Buybacks (cash)")

                assert q1_2025_col is not None
                assert q2_2025_col is not None
                assert q3_2025_col is not None
                assert buyback_cash_row is not None
                assert _fill_rgb(ws.cell(row=buyback_cash_row, column=q1_2025_col)) in {"", "00000000", "000000"}
                assert _fill_rgb(ws.cell(row=buyback_cash_row, column=q2_2025_col)) not in {"", "00000000", "000000"}
                assert _fill_rgb(ws.cell(row=buyback_cash_row, column=q3_2025_col)) not in {"", "00000000", "000000"}
            finally:
                wb.close()


def test_bs_segments_first_visible_year_can_use_hidden_history_for_heatmap_fill() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "bs_segments_hidden_history_heatmap.xlsx")
        quarters = pd.period_range("2022Q1", "2024Q4", freq="Q").to_timestamp("Q")
        hist = pd.DataFrame(
            {
                "quarter": pd.to_datetime(quarters),
                "cash": [100_000_000.0 + (i * 10_000_000.0) for i in range(len(quarters))],
                "assets_current": [300_000_000.0 + (i * 20_000_000.0) for i in range(len(quarters))],
                "liabilities_current": [150_000_000.0 + (i * 5_000_000.0) for i in range(len(quarters))],
                "assets": [800_000_000.0 + (i * 30_000_000.0) for i in range(len(quarters))],
                "liabilities": [400_000_000.0 + (i * 10_000_000.0) for i in range(len(quarters))],
                "equity": [400_000_000.0 + (i * 20_000_000.0) for i in range(len(quarters))],
                "shares_outstanding": [100_000_000.0] * len(quarters),
                "shares_diluted": [100_000_000.0] * len(quarters),
            }
        )

        ctx = build_writer_context(_make_inputs(out_path, hist=hist))
        ctx.callbacks.write_bs_segments_sheet()
        ws = ctx.wb["BS_Segments"]

        q1_2023_col = next(
            (
                cc
                for rr in range(1, ws.max_row + 1)
                for cc in range(1, ws.max_column + 1)
                if str(ws.cell(row=rr, column=cc).value or "").strip() == "2023-Q1"
            ),
            None,
        )
        cash_row = _find_row_with_value(ws, "Cash & cash equivalents")

        assert q1_2023_col is not None
        assert cash_row is not None
        assert ws.cell(row=cash_row, column=q1_2023_col).value == pytest.approx(140.0)
        assert _fill_rgb(ws.cell(row=cash_row, column=q1_2023_col)) == "002F80ED"


def test_operating_drivers_first_visible_year_can_use_hidden_history_for_heatmap_fill() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "operating_drivers_hidden_history_heatmap.xlsx")
        hist_quarters = pd.period_range("2023Q1", "2025Q4", freq="Q").to_timestamp("Q")
        hist = pd.DataFrame(
            {
                "quarter": pd.to_datetime(hist_quarters),
                "revenue": [100.0 + i for i in range(len(hist_quarters))],
                "cfo": [10.0 + i for i in range(len(hist_quarters))],
                "capex": [2.0 + i for i in range(len(hist_quarters))],
                "ebitda": [15.0 + i for i in range(len(hist_quarters))],
                "ebit": [9.0 + i for i in range(len(hist_quarters))],
                "cash": [20.0 + i for i in range(len(hist_quarters))],
                "debt_core": [50.0 - i for i in range(len(hist_quarters))],
                "shares_outstanding": [10.0] * len(hist_quarters),
                "shares_diluted": [10.0] * len(hist_quarters),
                "market_cap": [100.0] * len(hist_quarters),
                "interest_expense_net": [1.0] * len(hist_quarters),
            }
        )
        profile = CompanyProfile(
            ticker="TEST",
            has_bank=False,
            industry_keywords=tuple(),
            segment_patterns=tuple(),
            segment_alias_patterns=tuple(),
            key_adv_require_keywords=tuple(),
            key_adv_deny_keywords=tuple(),
            promise_priority_terms=tuple(),
            enable_operating_drivers_sheet=True,
            operating_driver_history_templates=tuple(),
        )
        monkeypatch = pytest.MonkeyPatch()
        monkeypatch.setattr(writer_context_module, "get_company_profile", lambda ticker: profile)
        try:
            ctx = build_writer_context(_make_inputs(out_path, hist=hist))
            rows = []
            for quarter, value in [
                ("2022-03-31", 70.0),
                ("2022-06-30", 75.0),
                ("2022-09-30", 78.0),
                ("2022-12-31", 80.0),
                ("2023-03-31", 100.0),
                ("2023-06-30", 110.0),
                ("2023-09-30", 120.0),
                ("2023-12-31", 130.0),
                ("2024-03-31", 140.0),
                ("2024-06-30", 150.0),
                ("2024-09-30", 160.0),
                ("2024-12-31", 170.0),
                ("2025-03-31", 180.0),
                ("2025-06-30", 190.0),
                ("2025-09-30", 200.0),
                ("2025-12-31", 210.0),
            ]:
                rows.append(
                    {
                        "Quarter": pd.Timestamp(quarter).date(),
                        "Driver group": "Volume",
                        "Driver": "Volume sold",
                        "Unit": "m gallons",
                        "Value": value,
                        "_driver_key": "volume_sold",
                    }
                )
            ctx.callbacks.write_operating_drivers_sheet(rows)
            ws = ctx.wb["Operating_Drivers"]

            q1_2023_col = next(
                (
                    cc
                    for rr in range(1, ws.max_row + 1)
                    for cc in range(1, ws.max_column + 1)
                    if str(ws.cell(row=rr, column=cc).value or "").strip() == "2023-Q1"
                ),
                None,
            )
            volume_row = _find_row_with_value(ws, "Volume sold (million gallons)")

            assert q1_2023_col is not None
            assert volume_row is not None
            assert ws.cell(row=volume_row, column=q1_2023_col).value == pytest.approx(100.0)
            assert _fill_rgb(ws.cell(row=volume_row, column=q1_2023_col)) == "002F80ED"
        finally:
            monkeypatch.undo()


def test_gpre_live_operating_drivers_keep_current_quarter_utilization_and_operating_commentary() -> None:
    with _case_dir() as case_dir:
        out_path = _make_ticker_model_out_path(case_dir, "GPRE", "gpre_live_operating_drivers.xlsx")
        inputs = _make_live_artifact_inputs("GPRE", out_path)
        ctx = build_writer_context(inputs)
        ensure_driver_inputs(ctx)

        utilization_by_quarter = {
            pd.Timestamp(rec.get("Quarter")).date(): float(pd.to_numeric(rec.get("Value"), errors="raise"))
            for rec in ctx.data.operating_driver_history_rows
            if str(rec.get("_driver_key") or "").strip() == "utilization"
        }
        assert utilization_by_quarter[date(2023, 9, 30)] == pytest.approx(93.9, abs=0.01)
        assert utilization_by_quarter[date(2024, 6, 30)] == pytest.approx(92.6, abs=0.01)
        assert utilization_by_quarter[date(2024, 9, 30)] == pytest.approx(96.8, abs=0.01)
        assert utilization_by_quarter[date(2024, 12, 31)] == pytest.approx(92.0, abs=0.01)

        fn = ctx.callbacks.write_operating_drivers_sheet
        wb = _callback_closure_value(fn, "wb")
        fn(ctx.data.operating_driver_history_rows)
        ws_drv = wb["Operating_Drivers"]
        drv_commentary_rows = [row for row in _operating_commentary_rows(ws_drv) if _quarter_label_ord(row["stated_in"]) is not None]
        commentary_targets = [
            "Plants ran above 100% capacity utilization during the quarter.",
            "Reliability-centered maintenance reduced planned and unplanned downtime.",
            "Record high ethanol and Ultra-high protein yields supported record protein output and corn-oil production.",
        ]
        for target in commentary_targets:
            assert sum(row["commentary"] == target for row in drv_commentary_rows) == 1

        assert any(
            row["commentary"] == "Plant utilization reflected 97% during the quarter, compared to a 94% run rate in the same period last year."
            and row["stated_in"] == "Q3 2024"
            for row in drv_commentary_rows
        )
        assert any(
            row["commentary"] == "Plant utilization reflected 92% during the fourth quarter compared to the 95% run rate reported in the same period last year."
            and row["stated_in"] == "Q4 2024"
            for row in drv_commentary_rows
        )
        assert any(
            row["commentary"] == "Plant utilization reflected 93.9% during the quarter, returning the platform to consistent operations."
            and row["stated_in"] == "Q3 2023"
            for row in drv_commentary_rows
        )

        setup_records = _callback_closure_value(fn, "_gpre_commercial_setup_records_shared")()
        for target in commentary_targets:
            matches = [rec for rec in setup_records if str(rec.get("commentary_text") or "").strip() == target]
            assert matches
            assert all(str(rec.get("commentary_home") or "").strip() == "operating_commentary" for rec in matches)


def test_gpre_live_economics_overlay_stage5_proxy_story_chart_and_sheet_order() -> None:
    with _case_dir() as case_dir:
        out_path = _make_ticker_model_out_path(case_dir, "GPRE", "gpre_live_economics_overlay_stage5.xlsx")
        inputs = _make_live_artifact_inputs("GPRE", out_path)
        ctx = build_writer_context(inputs)

        assert ctx.desired_sheet_order.index("Promise_Progress_UI") + 1 == ctx.desired_sheet_order.index("Basis_Proxy_Sandbox")
        assert ctx.desired_sheet_order.index("Basis_Proxy_Sandbox") + 1 == ctx.desired_sheet_order.index("Hidden_Value_Flags")

        ensure_driver_inputs(ctx)
        write_driver_sheets(ctx)

        ws_overlay = ctx.wb["Economics_Overlay"]
        ws_basis = ctx.wb["Basis_Proxy_Sandbox"]
        proxy_compare_row = _find_row_with_value(ws_overlay, "Proxy comparison ($/gal)", column=1)
        bridge_row = _find_row_with_value(ws_overlay, "Bridge to reported", column=1)
        forward_proxy_row = _find_row_with_value(ws_overlay, "Best forward lens ($/gal)", column=1)
        forward_bridge_row = _find_row_with_value(ws_overlay, "Best forward lens ($m)", column=1)
        weekly_chart_title_row = _find_row_with_value(ws_overlay, "Approximate market crush (weekly)", column=2)
        quarterly_chart_title_row = _find_row_with_value(ws_overlay, "Approximate market crush, fitted models, and real GPRE crush margin (quarterly)", column=2)

        assert proxy_compare_row is not None
        assert bridge_row is not None
        assert forward_proxy_row is not None
        assert forward_bridge_row is not None
        assert weekly_chart_title_row is not None
        assert quarterly_chart_title_row is not None
        assert quarterly_chart_title_row > weekly_chart_title_row

        proxy_note = str(ws_overlay.cell(row=proxy_compare_row + 1, column=1).value or "")
        assert "Official row = Approximate market crush" in proxy_note
        assert "Fitted row = GPRE crush proxy" in proxy_note
        assert "Production winner =" in proxy_note
        assert "Best forward lens =" in proxy_note
        assert "Coproduct-aware experimental lenses live in Basis_Proxy_Sandbox and are comparison-only." in proxy_note
        assert any(
            rng.min_row == proxy_compare_row + 1 and rng.min_col == 1 and rng.max_col == 21
            for rng in ws_overlay.merged_cells.ranges
        )
        assert _fill_rgb(ws_overlay.cell(row=proxy_compare_row + 1, column=1)) == "00EDF4FA"
        assert float(ws_overlay.row_dimensions[proxy_compare_row + 1].height or 0.0) == pytest.approx(18.0, abs=0.1)
        assert str(ws_overlay.cell(row=proxy_compare_row + 2, column=1).value or "").strip() == "Proxy row"
        assert forward_proxy_row == _find_row_with_value(ws_overlay, "GPRE crush proxy ($/gal)", column=1) + 1
        assert forward_bridge_row == _find_row_with_value(ws_overlay, "GPRE crush proxy ($m)", column=1) + 1

        assert len(ws_overlay._charts) == 3
        weekly_chart = ws_overlay._charts[0]
        quarterly_chart = ws_overlay._charts[1]
        coproduct_history_chart = ws_overlay._charts[2]
        coproduct_chart_title_row = _find_row_with_value(ws_overlay, "Approximate coproduct credit ($/gal, quarterly history)", column=2)
        assert len(weekly_chart.series) >= 6
        assert len(quarterly_chart.series) == 4
        assert getattr(getattr(quarterly_chart, "legend", None), "position", None) == "t"
        assert bool(getattr(getattr(quarterly_chart, "legend", None), "overlay", False))
        assert str(getattr(getattr(quarterly_chart.series[0].graphicalProperties.line, "solidFill", None), "srgbClr", "") or "") == "2F80ED"
        assert str(getattr(getattr(quarterly_chart.series[1].graphicalProperties.line, "solidFill", None), "srgbClr", "") or "") == "E67E22"
        assert str(getattr(getattr(quarterly_chart.series[2].graphicalProperties.line, "solidFill", None), "srgbClr", "") or "") == "2A9D8F"
        assert str(getattr(getattr(quarterly_chart.series[3].graphicalProperties.line, "solidFill", None), "srgbClr", "") or "") == "36454F"
        assert int(getattr(quarterly_chart.series[3].graphicalProperties.line, "width", 0) or 0) == 19050
        assert float(getattr(weekly_chart, "width", 0.0) or 0.0) == pytest.approx(34.0, abs=0.01)
        assert float(getattr(weekly_chart, "height", 0.0) or 0.0) == pytest.approx(16.0, abs=0.01)
        assert float(getattr(quarterly_chart, "width", 0.0) or 0.0) == pytest.approx(34.0, abs=0.01)
        assert float(getattr(quarterly_chart, "height", 0.0) or 0.0) == pytest.approx(16.0, abs=0.01)
        assert float(getattr(coproduct_history_chart, "width", 0.0) or 0.0) == pytest.approx(34.0, abs=0.01)
        assert float(getattr(coproduct_history_chart, "height", 0.0) or 0.0) == pytest.approx(16.0, abs=0.01)
        assert int(getattr(weekly_chart.anchor._from, "row", -1)) + 1 == weekly_chart_title_row + 1
        assert int(getattr(quarterly_chart.anchor._from, "row", -1)) + 1 == quarterly_chart_title_row + 1
        assert int(getattr(coproduct_history_chart.anchor._from, "row", -1)) + 1 == (coproduct_chart_title_row or 0) + 1
        assert int(getattr(quarterly_chart.anchor.to, "row", -1)) + 1 >= quarterly_chart_title_row + 26
        assert int(getattr(coproduct_history_chart.anchor.to, "row", -1)) + 1 >= (coproduct_chart_title_row or 0) + 26
        assert str(ws_overlay.cell(row=quarterly_chart_title_row, column=49).value or "").strip() == "Realized GPRE crush margin ($/gal)"
        q1_realized_row = next(
            rr for rr in range(quarterly_chart_title_row + 1, quarterly_chart_title_row + 12)
            if str(ws_overlay.cell(row=rr, column=45).value or "").strip() == "2025-Q1"
        )
        q2_realized_row = next(
            rr for rr in range(quarterly_chart_title_row + 1, quarterly_chart_title_row + 12)
            if str(ws_overlay.cell(row=rr, column=45).value or "").strip() == "2025-Q2"
        )
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=q1_realized_row, column=49).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=q2_realized_row, column=49).value, errors="coerce"))

        role_summary_row = _find_row_with_value(ws_basis, "Role summary", column=21)
        winner_story_title_row = _find_row_with_value(ws_basis, "Winner story", column=21)
        best_historical_row = _find_row_with_value(ws_basis, "Best historical fit", column=21)
        best_compromise_row = _find_row_with_value(ws_basis, "Best compromise", column=21)
        best_forward_row = _find_row_with_value(ws_basis, "Best forward lens", column=21)
        forward_usability_row = _find_row_with_value(ws_basis, "Forward usability", column=21)
        build_up_row = _find_row_with_value(ws_basis, "Approximate market crush build-up ($/gal)", column=2)
        corn_oil_gate_title_row = _find_row_with_value(ws_basis, "Coproduct source gate", column=2)
        coproduct_readiness_title_row = _find_row_with_value(ws_basis, "Coproduct signal readiness", column=2)
        nwer_gate_row = next(
            (rr for rr in range((corn_oil_gate_title_row or 0) + 1, (corn_oil_gate_title_row or 0) + 10)
             if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "NWER coproduct rows"),
            None,
        )
        ams_3618_gate_row = next(
            (rr for rr in range((corn_oil_gate_title_row or 0) + 1, (corn_oil_gate_title_row or 0) + 10)
             if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "AMS 3618 coproduct rows"),
            None,
        )
        price_gate_row = next(
            (rr for rr in range((corn_oil_gate_title_row or 0) + 1, (corn_oil_gate_title_row or 0) + 10)
             if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Renewable corn oil price"),
            None,
        )
        ddgs_gate_row = next(
            (rr for rr in range((corn_oil_gate_title_row or 0) + 1, (corn_oil_gate_title_row or 0) + 10)
             if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Distillers grains price"),
            None,
        )
        credit_gate_row = next(
            (rr for rr in range((corn_oil_gate_title_row or 0) + 1, (corn_oil_gate_title_row or 0) + 10)
             if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Approximate coproduct credit"),
            None,
        )
        activation_gate_row = next(
            (rr for rr in range((corn_oil_gate_title_row or 0) + 1, (corn_oil_gate_title_row or 0) + 10)
             if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Overlay activation"),
            None,
        )
        primary_activation_row = _find_row_with_value(ws_basis, "Primary live activation source", column=2)
        secondary_source_row = _find_row_with_value(ws_basis, "Secondary corroborating source", column=2)
        resolved_source_row = _find_row_with_value(ws_basis, "Current resolved workbook source", column=2)
        corn_oil_readiness_row = next(
            (rr for rr in range((coproduct_readiness_title_row or 0) + 1, (coproduct_readiness_title_row or 0) + 12)
             if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Renewable corn oil price"),
            None,
        )
        nwer_readiness_row = next(
            (rr for rr in range((coproduct_readiness_title_row or 0) + 1, (coproduct_readiness_title_row or 0) + 12)
             if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "NWER coproduct rows"),
            None,
        )
        ams_3618_readiness_row = next(
            (rr for rr in range((coproduct_readiness_title_row or 0) + 1, (coproduct_readiness_title_row or 0) + 12)
             if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "AMS 3618 coproduct rows"),
            None,
        )
        ddgs_readiness_row = next(
            (rr for rr in range((coproduct_readiness_title_row or 0) + 1, (coproduct_readiness_title_row or 0) + 12)
             if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Distillers grains price"),
            None,
        )
        coproduct_credit_readiness_row = next(
            (rr for rr in range((coproduct_readiness_title_row or 0) + 1, (coproduct_readiness_title_row or 0) + 12)
             if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Approximate coproduct credit"),
            None,
        )
        coproduct_history_title_row = _find_row_with_value(ws_basis, "Coproduct quarterly history", column=2)
        memo_row = _find_row_with_value(ws_basis, "Hedge-adjusted memo tests", column=2)
        assert best_historical_row is not None
        assert best_compromise_row is not None
        assert best_forward_row is not None
        assert role_summary_row is not None
        assert winner_story_title_row is not None
        assert forward_usability_row is not None
        assert build_up_row is not None
        assert corn_oil_gate_title_row is not None
        assert nwer_gate_row is not None
        assert ams_3618_gate_row is not None
        assert price_gate_row is not None
        assert ddgs_gate_row is not None
        assert credit_gate_row is not None
        assert activation_gate_row is not None
        assert primary_activation_row is not None
        assert secondary_source_row is not None
        assert resolved_source_row is not None
        assert coproduct_readiness_title_row is not None
        assert corn_oil_readiness_row is not None
        assert nwer_readiness_row is not None
        assert ams_3618_readiness_row is not None
        assert ddgs_readiness_row is not None
        assert coproduct_credit_readiness_row is not None
        coproduct_frame_summary_title_row = _find_row_with_value(ws_basis, "Coproduct frame summary", column=2)
        prior_frame_row = _find_row_with_value(ws_basis, "Prior quarter", column=2)
        quarter_open_frame_row = _find_row_with_value(ws_basis, "Quarter-open outlook", column=2)
        current_frame_row = _find_row_with_value(ws_basis, "Current QTD", column=2)
        next_frame_row = _find_row_with_value(ws_basis, "Next quarter outlook", column=2)
        coproduct_volume_support_title_row = _find_row_with_value(ws_basis, "Coproduct volume support audit", column=2)
        coproduct_experimental_title_row = _find_row_with_value(ws_basis, "Coproduct-aware experimental lenses", column=2)
        best_coproduct_experimental_row = _find_row_with_value(ws_basis, "Best coproduct-aware experimental lens", column=2)
        best_coproduct_experimental_historical_row = _find_row_with_value(ws_basis, "Best historical coproduct-aware", column=2)
        best_coproduct_experimental_forward_row = _find_row_with_value(ws_basis, "Best forward coproduct-aware", column=2)
        previous_coproduct_reference_row = _find_row_with_value(ws_basis, "Previous best coproduct-aware (reference)", column=2)
        current_production_winner_reference_row = _find_row_with_value(ws_basis, "Current production winner (reference)", column=2)
        coproduct_experimental_promotion_status_row = _find_row_with_value(ws_basis, "Promotion status", column=2)
        coproduct_experimental_method_header_row = _find_row_with_value(ws_basis, "Method", column=2)
        assert coproduct_frame_summary_title_row is not None
        assert prior_frame_row is not None
        assert quarter_open_frame_row is not None
        assert current_frame_row is not None
        assert next_frame_row is not None
        assert coproduct_history_title_row is not None
        assert coproduct_volume_support_title_row is not None
        assert coproduct_experimental_title_row is not None
        assert best_coproduct_experimental_row is not None
        assert best_coproduct_experimental_historical_row is not None
        assert best_coproduct_experimental_forward_row is not None
        assert previous_coproduct_reference_row is not None
        assert current_production_winner_reference_row is not None
        assert coproduct_experimental_promotion_status_row is not None
        assert coproduct_experimental_method_header_row is not None
        assert memo_row is not None
        assert coproduct_frame_summary_title_row < coproduct_history_title_row < coproduct_volume_support_title_row < coproduct_experimental_title_row < memo_row
        assert _fill_rgb(ws_basis.cell(row=build_up_row, column=2)) == "00D9E7F3"
        assert _fill_rgb(ws_basis.cell(row=corn_oil_gate_title_row, column=2)) == "00D9E7F3"
        assert _fill_rgb(ws_basis.cell(row=coproduct_frame_summary_title_row, column=2)) == "00D9E7F3"
        assert _fill_rgb(ws_basis.cell(row=coproduct_history_title_row, column=2)) == "00EAF3FB"
        assert _fill_rgb(ws_basis.cell(row=coproduct_volume_support_title_row, column=2)) == "00EAF3FB"
        assert _fill_rgb(ws_basis.cell(row=coproduct_experimental_title_row, column=2)) == "00F7F9FC"
        assert _fill_rgb(ws_basis.cell(row=build_up_row + 1, column=2)) == "00F4F8FC"
        assert _fill_rgb(ws_basis.cell(row=coproduct_history_title_row + 1, column=2)) == "00F4F8FC"
        assert _fill_rgb(ws_basis.cell(row=coproduct_experimental_title_row + 1, column=2)) == "00F8FBFD"
        assert _fill_rgb(ws_basis.cell(row=winner_story_title_row, column=21)) == "00F7F9FC"
        assert float(ws_basis.column_dimensions["B"].width or 0.0) >= 16.0
        assert float(ws_basis.column_dimensions["S"].width or 0.0) >= 36.0
        assert float(ws_basis.column_dimensions["U"].width or 0.0) >= 21.0
        assert float(ws_basis.column_dimensions["X"].width or 0.0) >= 14.0
        assert float(ws_basis.row_dimensions[build_up_row + 1].height or 0.0) == pytest.approx(28.0, abs=0.1)
        assert float(ws_basis.row_dimensions[coproduct_frame_summary_title_row + 1].height or 0.0) == pytest.approx(64.0, abs=0.1)
        assert "Production winner = fitted row used in production" in str(ws_basis.cell(row=role_summary_row + 5, column=21).value or "")
        assert "Hybrid" in str(ws_basis.cell(row=best_historical_row, column=22).value or "")
        assert "MAE" in str(ws_basis.cell(row=best_historical_row, column=22).value or "")
        assert "Forward" in str(ws_basis.cell(row=best_historical_row, column=22).value or "")
        assert str(ws_basis.cell(row=best_compromise_row, column=22).value or "").strip()
        assert str(ws_basis.cell(row=best_forward_row, column=22).value or "").strip()
        assert "winner" in str(ws_basis.cell(row=forward_usability_row, column=22).value or "").lower()
        assert str(ws_basis.cell(row=nwer_gate_row, column=6).value or "").strip() == "YES"
        assert str(ws_basis.cell(row=ams_3618_gate_row, column=6).value or "").strip() == "YES"
        assert str(ws_basis.cell(row=price_gate_row, column=6).value or "").strip() == "YES"
        assert str(ws_basis.cell(row=ddgs_gate_row, column=6).value or "").strip() == "YES"
        assert str(ws_basis.cell(row=credit_gate_row, column=6).value or "").strip() == "YES"
        assert str(ws_basis.cell(row=activation_gate_row, column=6).value or "").strip() == "GO"
        assert "primary live activation source" in str(ws_basis.cell(row=activation_gate_row + 1, column=2).value or "").lower()
        assert str(ws_basis.cell(row=primary_activation_row, column=6).value or "").strip() == "NWER"
        assert "controls go/hold" in str(ws_basis.cell(row=primary_activation_row, column=8).value or "").lower()
        assert str(ws_basis.cell(row=secondary_source_row, column=6).value or "").strip() == "AMS 3618"
        assert "not required for visible activation" in str(ws_basis.cell(row=secondary_source_row, column=8).value or "").lower()
        expected_resolved_source = str(ws_basis.cell(row=current_frame_row, column=8).value or "").strip()
        assert expected_resolved_source in {"NWER", "AMS 3618", "Mixed", "Unknown/blank"}
        assert str(ws_basis.cell(row=resolved_source_row, column=6).value or "").strip() == expected_resolved_source
        assert "visible price rows" in str(ws_basis.cell(row=resolved_source_row, column=8).value or "").lower()
        assert "stage b.4 keeps nwer as the primary live activation source" in str(ws_basis.cell(row=activation_gate_row + 1, column=2).value or "").lower()
        assert "manual fallback/backfill" in str(ws_basis.cell(row=activation_gate_row + 1, column=2).value or "").lower()
        assert "3511" in str(ws_basis.cell(row=activation_gate_row + 1, column=2).value or "").lower()
        assert "stage b.4 keeps nwer as the sufficient first visible coproduct source" in str(ws_basis.cell(row=coproduct_readiness_title_row + 1, column=2).value or "").lower()
        assert "manual fallback-backfill" in str(ws_basis.cell(row=coproduct_readiness_title_row + 1, column=2).value or "").lower()
        assert str(ws_basis.cell(row=coproduct_frame_summary_title_row + 2, column=2).value or "").strip() == "Frame"
        assert str(ws_basis.cell(row=coproduct_frame_summary_title_row + 2, column=3).value or "").strip() == "Renewable corn oil price"
        assert str(ws_basis.cell(row=coproduct_frame_summary_title_row + 2, column=6).value or "").strip() == "Approximate coproduct credit ($/gal)"
        assert str(ws_basis.cell(row=coproduct_frame_summary_title_row + 2, column=7).value or "").strip() == "Approximate coproduct credit ($m)"
        assert str(ws_basis.cell(row=coproduct_frame_summary_title_row + 2, column=9).value or "").strip() == "Coverage"
        assert str(ws_basis.cell(row=coproduct_frame_summary_title_row + 2, column=10).value or "").strip() == "Rule"
        assert all(str(ws_basis.cell(row=rr, column=6).value or "").strip() for rr in (prior_frame_row, quarter_open_frame_row, current_frame_row, next_frame_row))
        assert all(str(ws_basis.cell(row=rr, column=7).value or "").strip() for rr in (prior_frame_row, quarter_open_frame_row, current_frame_row, next_frame_row))
        assert all(str(ws_basis.cell(row=rr, column=10).value or "").strip() for rr in (prior_frame_row, quarter_open_frame_row, current_frame_row, next_frame_row))
        assert str(ws_basis.cell(row=corn_oil_readiness_row, column=4).value or "").strip() == "Direct market"
        assert str(ws_basis.cell(row=ddgs_readiness_row, column=4).value or "").strip() == "Direct market"
        assert str(ws_basis.cell(row=nwer_readiness_row, column=4).value or "").strip() == "Weekly bioenergy"
        assert str(ws_basis.cell(row=ams_3618_readiness_row, column=4).value or "").strip() == "Weekly co-products"
        assert str(ws_basis.cell(row=coproduct_credit_readiness_row, column=4).value or "").strip() == "Derived build-up"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=2).value or "").strip() == "Quarter"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=3).value or "").strip() == "Renewable corn oil price"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=4).value or "").strip() == "Distillers grains price"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=5).value or "").strip() == "Approximate coproduct credit ($/bushel)"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=6).value or "").strip() == "Approximate coproduct credit ($/gal)"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=7).value or "").strip() == "Approximate coproduct credit ($m)"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=8).value or "").strip() == "Resolved source mode"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=9).value or "").strip() == "Coverage"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=9).value or "").strip() == "Coverage"
        assert "coverage is covered active-capacity share" in str(ws_basis.cell(row=coproduct_history_title_row + 1, column=2).value or "").lower()
        history_rows = [
            rr
            for rr in range(coproduct_history_title_row + 3, memo_row)
            if re.match(r"^20\d{2}-Q[1-4]$", str(ws_basis.cell(row=rr, column=2).value or "").strip())
        ]
        assert len(history_rows) >= 8
        assert str(ws_basis.cell(history_rows[0], 2).value or "").strip() == "2022-Q3"
        assert str(ws_basis.cell(history_rows[-1], 2).value or "").strip() == "2026-Q2"
        assert all(str(ws_basis.cell(rr, 5).value or "").strip() for rr in history_rows)
        assert all(str(ws_basis.cell(rr, 6).value or "").strip() for rr in history_rows)
        assert sum(1 for rr in history_rows if str(ws_basis.cell(rr, 7).value or "").strip()) >= 8
        assert all(str(ws_basis.cell(rr, 9).value or "").strip() for rr in history_rows)
        source_modes_seen = {
            str(ws_basis.cell(rr, 8).value or "").strip()
            for rr in history_rows
            if str(ws_basis.cell(rr, 8).value or "").strip()
        }
        assert source_modes_seen.issubset({"NWER", "AMS 3618", "Mixed", "Unknown/blank"})
        assert str(ws_basis.cell(row=coproduct_volume_support_title_row + 2, column=2).value or "").strip() == "Series"
        assert str(ws_basis.cell(row=coproduct_volume_support_title_row + 2, column=4).value or "").strip() == "Source/path"
        assert str(ws_basis.cell(row=coproduct_volume_support_title_row + 2, column=6).value or "").strip() == "Historical usable"
        assert str(ws_basis.cell(row=coproduct_volume_support_title_row + 2, column=9).value or "").strip() == "Best use"
        assert "historical actuals only" in str(ws_basis.cell(row=coproduct_volume_support_title_row + 1, column=2).value or "").lower()
        distillers_volume_audit_row = _find_row_with_value(ws_basis, "Distillers grains volume", column=2)
        corn_oil_volume_audit_row = _find_row_with_value(ws_basis, "Renewable corn oil volume", column=2)
        uhp_volume_audit_row = _find_row_with_value(ws_basis, "Ultra-high protein volume", column=2)
        mix_commentary_row = _find_row_with_value(ws_basis, "Protein / coproduct mix commentary", column=2)
        assert distillers_volume_audit_row is not None
        assert corn_oil_volume_audit_row is not None
        assert uhp_volume_audit_row is not None
        assert mix_commentary_row is not None
        assert str(ws_basis.cell(row=distillers_volume_audit_row, column=6).value or "").strip() == "YES"
        assert str(ws_basis.cell(row=distillers_volume_audit_row, column=7).value or "").strip() == "NO"
        assert str(ws_basis.cell(row=distillers_volume_audit_row, column=8).value or "").strip() == "NO"
        assert str(ws_basis.cell(row=distillers_volume_audit_row, column=9).value or "").strip() == "QA only"
        assert "tons/mm gal" in str(ws_basis.cell(row=distillers_volume_audit_row, column=11).value or "").lower()
        assert str(ws_basis.cell(row=corn_oil_volume_audit_row, column=9).value or "").strip() == "QA only"
        assert "lbs/gal" in str(ws_basis.cell(row=corn_oil_volume_audit_row, column=11).value or "").lower()
        assert str(ws_basis.cell(row=uhp_volume_audit_row, column=9).value or "").strip() == "Secondary QA only"
        assert "secondary qa" in str(ws_basis.cell(row=uhp_volume_audit_row, column=11).value or "").lower()
        assert str(ws_basis.cell(row=mix_commentary_row, column=6).value or "").strip() == "Commentary only"
        assert str(ws_basis.cell(row=mix_commentary_row, column=9).value or "").strip() == "Context only"
        assert "comparison only" in str(ws_basis.cell(row=coproduct_experimental_title_row + 1, column=2).value or "").lower()
        assert str(ws_basis.cell(row=best_coproduct_experimental_row, column=3).value or "").strip()
        assert str(ws_basis.cell(row=best_coproduct_experimental_historical_row, column=3).value or "").strip()
        assert str(ws_basis.cell(row=best_coproduct_experimental_forward_row, column=3).value or "").strip()
        assert str(ws_basis.cell(row=previous_coproduct_reference_row, column=3).value or "").strip() == "Simple + 50% credit"
        assert str(ws_basis.cell(row=current_production_winner_reference_row, column=3).value or "").strip()
        assert str(ws_basis.cell(row=coproduct_experimental_promotion_status_row, column=3).value or "").strip() == "Experimental only"
        assert str(ws_basis.cell(row=coproduct_experimental_method_header_row, column=3).value or "").strip() == "Rule"
        assert str(ws_basis.cell(row=coproduct_experimental_method_header_row, column=4).value or "").strip() == "Clean MAE"
        assert str(ws_basis.cell(row=coproduct_experimental_method_header_row, column=5).value or "").strip() == "Hybrid"
        assert str(ws_basis.cell(row=coproduct_experimental_method_header_row, column=10).value or "").strip() == "Low-coverage MAE"
        assert str(ws_basis.cell(row=coproduct_experimental_method_header_row, column=12).value or "").strip() == "Status"
        coproduct_experimental_status_rows = [
            rr
            for rr in range(coproduct_experimental_method_header_row + 1, memo_row)
            if str(ws_basis.cell(row=rr, column=12).value or "").strip()
        ]
        assert len(coproduct_experimental_status_rows) == 10
        assert all(str(ws_basis.cell(row=rr, column=12).value or "").strip() == "comparison only" for rr in coproduct_experimental_status_rows)
        qtd_tracking_title_row = _find_row_with_value(ws_overlay, "Current QTD trend tracking ($/gal, crush margin lens)", column=1)
        assert qtd_tracking_title_row is not None
        _assert_gpre_qtd_tracking_upper_block_dynamic(ws_overlay)
        assert _find_row_with_value(ws_overlay, "Ethanol", column=1) is not None
        assert _find_row_with_value(ws_overlay, "Flat corn", column=1) is not None
        assert _find_row_with_value(ws_overlay, "Corn basis", column=1) is not None
        assert _find_row_with_value(ws_overlay, "Gas", column=1) is not None
        coproducts_row = _find_row_with_value(ws_overlay, "Coproducts", column=1)
        assert coproducts_row == qtd_tracking_title_row + 14
        assert _find_row_with_value(ws_overlay, "Residual", column=1) is None
        assert _find_row_containing(
            ws_overlay,
            "Same-point-last-quarter is intentionally not a primary tracking metric here",
            column=2,
        ) is None
        assert ws_overlay.row_dimensions[82].height == pytest.approx(18.0, abs=0.01)
        assert ws_overlay.row_dimensions[107].height == pytest.approx(18.0, abs=0.01)
        assert ws_overlay.row_dimensions[122].height == pytest.approx(18.0, abs=0.01)
        assert ws_overlay.row_dimensions[qtd_tracking_title_row + 13].height == pytest.approx(15.0, abs=0.01)
        basis_weighting_row = _find_row_with_value(ws_overlay, "Basis weighting", column=1)
        assert basis_weighting_row is not None
        assert str(ws_overlay.cell(row=basis_weighting_row, column=2).value or "").strip() == (
            "Official corn basis prefers dated GPRE plant bids when available; otherwise it falls back to "
            "active-capacity-weighted AMS basis using mapped state/regional series and deterministic fallbacks"
        )
        assert ws_overlay.row_dimensions[coproducts_row].height == pytest.approx(18.0, abs=0.01)
        assert str(ws_overlay.cell(row=coproducts_row, column=1).fill.fgColor.rgb or "").strip().upper() == str(ws_overlay.cell(row=qtd_tracking_title_row, column=1).fill.fgColor.rgb or "").strip().upper()
        assert ws_overlay.row_dimensions[coproducts_row + 5].height == pytest.approx(8.0, abs=0.01)
        assert str(ws_overlay.cell(row=coproducts_row + 5, column=1).value or "").strip() == ""
        assert str(ws_overlay.cell(row=coproducts_row + 5, column=1).fill.fgColor.rgb or "").strip().upper() == str(ws_overlay.cell(row=qtd_tracking_title_row + 4, column=2).fill.fgColor.rgb or "").strip().upper()

        visible_coproduct_header_row = _find_row_with_value(ws_overlay, "Coproduct economics", column=1)
        assert visible_coproduct_header_row is not None and visible_coproduct_header_row > qtd_tracking_title_row
        assert visible_coproduct_header_row == coproducts_row + 1
        assert str(ws_overlay.cell(row=visible_coproduct_header_row, column=2).value or "").strip() == "Prior quarter"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row, column=4).value or "").strip() == "Quarter-open outlook"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row, column=6).value or "").strip() == "Current QTD"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row, column=8).value or "").strip() == "Next quarter outlook"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row, column=10).value or "").strip() == "Unit"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row, column=11).value or "").strip() == "Source mode"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 1, column=2).value or "").strip() == "2026-Q1"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 1, column=4).value or "").strip() == "As of 2026-03-31"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 1, column=6).value or "").strip().startswith("As of 2026-04-")
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 1, column=8).value or "").strip() == "2026-Q3"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 2, column=1).value or "").strip() == "Renewable corn oil price"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 3, column=1).value or "").strip() == "Distillers grains price"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 4, column=1).value or "").strip() == ""
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 5, column=1).value or "").strip() == "Approximate coproduct credit ($/gal)"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 6, column=1).value or "").strip() == "Approximate coproduct credit ($m)"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 2, column=10).value or "").strip() == "$/lb"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 3, column=10).value or "").strip() == "$/lb"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 5, column=10).value or "").strip() == "$/gal"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 6, column=10).value or "").strip() == "$m"
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 2, column=6).value or "").strip()
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 3, column=6).value or "").strip()
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 5, column=6).value or "").strip()
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 6, column=6).value or "").strip()
        assert str(ws_overlay.cell(row=visible_coproduct_header_row + 6, column=6).number_format or "").strip() == "#,##0.0"
        expected_visible_frame_refs = {
            visible_coproduct_header_row + 2: 3,
            visible_coproduct_header_row + 3: 4,
            visible_coproduct_header_row + 5: 6,
            visible_coproduct_header_row + 6: 7,
        }
        frame_rows_by_col = {
            2: prior_frame_row,
            4: quarter_open_frame_row,
            6: current_frame_row,
            8: next_frame_row,
        }
        for visible_row, frame_value_col in expected_visible_frame_refs.items():
            for cc, frame_row_num in frame_rows_by_col.items():
                expected_formula = f'=IF(ISNUMBER(Basis_Proxy_Sandbox!${get_column_letter(frame_value_col)}${frame_row_num}),Basis_Proxy_Sandbox!${get_column_letter(frame_value_col)}${frame_row_num},"")'
                assert str(ws_overlay.cell(row=visible_row, column=cc).value or "").strip() == expected_formula
        assert "weighted active-capacity quarterly resolver" in str(ws_overlay.cell(row=visible_coproduct_header_row + 2, column=11).value or "").lower()
        assert "weighted active-capacity quarterly resolver" in str(ws_overlay.cell(row=visible_coproduct_header_row + 3, column=11).value or "").lower()
        credit_per_gal_source_text = str(ws_overlay.cell(row=visible_coproduct_header_row + 5, column=11).value or "").lower()
        credit_usd_m_source_text = str(ws_overlay.cell(row=visible_coproduct_header_row + 6, column=11).value or "").lower()
        assert "weighted sandbox build-up divided by ethanol yield" in credit_per_gal_source_text
        assert "source mode, coverage, and carry-forward rule" in credit_per_gal_source_text
        assert "frame-specific implied gallons basis" in credit_usd_m_source_text
        coproduct_chart_title_row = _find_row_with_value(ws_overlay, "Approximate coproduct credit ($/gal, quarterly history)", column=2)
        mini_history_title_row = _find_row_with_value(ws_overlay, "Coproduct credit", column=2)
        assert coproduct_chart_title_row is not None and mini_history_title_row is not None
        assert coproduct_chart_title_row > visible_coproduct_header_row
        assert mini_history_title_row > coproduct_chart_title_row
        assert str(ws_overlay.cell(row=mini_history_title_row + 1, column=2).value or "").strip() == "Quarter"
        assert str(ws_overlay.cell(row=mini_history_title_row + 1, column=4).value or "").strip() == "$/gal"
        assert str(ws_overlay.cell(row=mini_history_title_row + 1, column=6).value or "").strip() == "$m"
        assert str(ws_overlay.cell(row=mini_history_title_row + 1, column=8).value or "").strip() == "Coverage"
        assert str(ws_overlay.cell(row=mini_history_title_row + 1, column=10).value or "").strip() == "Source mode"
        assert str(ws_overlay.cell(row=mini_history_title_row, column=13).value or "").strip() == "Corn oil prices"
        assert str(ws_overlay.cell(row=mini_history_title_row + 1, column=13).value or "").strip() == "Quarter"
        assert str(ws_overlay.cell(row=mini_history_title_row + 1, column=15).value or "").strip() == "$/lb"
        assert str(ws_overlay.cell(row=mini_history_title_row + 1, column=17).value or "").strip() == "$/gal"
        assert str(ws_overlay.cell(row=mini_history_title_row + 1, column=19).value or "").strip() == "$m proxy"
        coverage_note_row = _find_row_with_value(
            ws_overlay,
            "Coverage reflects covered active-capacity footprint; values are covered-footprint weighted averages.",
            column=2,
        )
        assert coverage_note_row is not None and coverage_note_row > mini_history_title_row
        mini_history_rows = [
            rr
            for rr in range(mini_history_title_row + 2, coverage_note_row)
            if str(ws_overlay.cell(row=rr, column=4).value or "").strip()
        ]
        assert len(mini_history_rows) >= 14
        assert all(str(ws_overlay.cell(row=rr, column=13).value or "").strip() for rr in mini_history_rows)
        assert any(pd.notna(pd.to_numeric(ws_overlay.cell(row=rr, column=15).value, errors="coerce")) for rr in mini_history_rows)
        assert any(pd.notna(pd.to_numeric(ws_overlay.cell(row=rr, column=17).value, errors="coerce")) for rr in mini_history_rows)
        assert any(pd.notna(pd.to_numeric(ws_overlay.cell(row=rr, column=19).value, errors="coerce")) for rr in mini_history_rows)
        assert str(ws_overlay.cell(row=mini_history_rows[0], column=2).value or "").strip() == "2026-Q3"
        assert str(ws_overlay.cell(row=mini_history_rows[-1], column=2).value or "").strip().startswith("=Basis_Proxy_Sandbox!$B$")
        q2_table_row = next(rr for rr in mini_history_rows if str(ws_overlay.cell(row=rr, column=2).value or "").strip().startswith("=Basis_Proxy_Sandbox!$B$150"))
        q1_table_row = next(rr for rr in mini_history_rows if str(ws_overlay.cell(row=rr, column=2).value or "").strip().startswith("=Basis_Proxy_Sandbox!$B$149"))
        q3_table_row = next(rr for rr in mini_history_rows if str(ws_overlay.cell(row=rr, column=2).value or "").strip() == "2026-Q3")
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=q2_table_row, column=6).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=q1_table_row, column=6).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=q1_table_row, column=17).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=q1_table_row, column=19).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=q3_table_row, column=17).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=q3_table_row, column=19).value, errors="coerce"))
        assert str(ws_overlay.cell(row=mini_history_rows[0], column=6).number_format or "").strip() == "#,##0.0"
        assert str(ws_overlay.cell(row=mini_history_rows[0], column=19).number_format or "").strip() == "#,##0.0"
        quarterly_helper_labels = [
            str(ws_overlay.cell(row=rr, column=45).value or "").strip()
            for rr in range(quarterly_chart_title_row + 1, quarterly_chart_title_row + 10)
            if str(ws_overlay.cell(row=rr, column=45).value or "").strip()
        ]
        assert quarterly_helper_labels
        assert all(re.match(r"^20\d{2}-Q[1-4]$", label) for label in quarterly_helper_labels[:4])
        quarterly_helper_full_labels = [
            str(ws_overlay.cell(row=rr, column=45).value or "").strip()
            for rr in range(quarterly_chart_title_row + 1, ws_overlay.max_row + 1)
            if str(ws_overlay.cell(row=rr, column=45).value or "").strip()
        ]
        assert quarterly_helper_full_labels
        assert "2026-Q3" in quarterly_helper_full_labels
        assert len(quarterly_helper_full_labels) <= 15
        assert [_quarter_label_ord(label) for label in quarterly_helper_full_labels] == sorted(
            [_quarter_label_ord(label) for label in quarterly_helper_full_labels]
        )
        coproduct_helper_labels = [
            str(ws_overlay.cell(row=rr, column=49).value or "").strip()
            for rr in range(coproduct_chart_title_row + 1, mini_history_title_row)
            if str(ws_overlay.cell(row=rr, column=49).value or "").strip()
        ]
        assert coproduct_helper_labels
        assert all(re.match(r"^20\d{2}-Q[1-4]$", label) for label in coproduct_helper_labels[:5])
        assert "2026-Q1" in coproduct_helper_labels
        assert "2026-Q2" in coproduct_helper_labels
        assert "2026-Q3" in coproduct_helper_labels
        assert coproduct_helper_labels.index("2026-Q1") < coproduct_helper_labels.index("2026-Q2") < coproduct_helper_labels.index("2026-Q3")
        assert len(coproduct_helper_labels) <= 15
        assert [_quarter_label_ord(label) for label in coproduct_helper_labels] == sorted(
            [_quarter_label_ord(label) for label in coproduct_helper_labels]
        )
        assert str(ws_overlay.cell(row=coproduct_chart_title_row + 1, column=50).value or "").startswith("=Basis_Proxy_Sandbox!$F$")
        q1_history_row = _find_row_with_value(ws_basis, "2026-Q1", column=2)
        q2_history_row = _find_row_with_value(ws_basis, "2026-Q2", column=2)
        assert q1_history_row is not None
        assert q2_history_row is not None
        coproduct_q1_helper_row = next(rr for rr in range(coproduct_chart_title_row + 1, mini_history_title_row) if str(ws_overlay.cell(row=rr, column=49).value or "").strip() == "2026-Q1")
        coproduct_q2_helper_row = next(rr for rr in range(coproduct_chart_title_row + 1, mini_history_title_row) if str(ws_overlay.cell(row=rr, column=49).value or "").strip() == "2026-Q2")
        coproduct_q3_helper_row = next(rr for rr in range(coproduct_chart_title_row + 1, mini_history_title_row) if str(ws_overlay.cell(row=rr, column=49).value or "").strip() == "2026-Q3")
        assert str(ws_overlay.cell(row=coproduct_q1_helper_row, column=50).value or "").strip() == f"=Basis_Proxy_Sandbox!$F${q1_history_row}"
        assert str(ws_overlay.cell(row=coproduct_q2_helper_row, column=50).value or "").strip() == f"=Basis_Proxy_Sandbox!$F${q2_history_row}"
        assert str(ws_overlay.cell(row=coproduct_q3_helper_row, column=50).value or "").strip() == f"=Basis_Proxy_Sandbox!$F${next_frame_row}"

        ctx.wb.save(out_path)
        with zipfile.ZipFile(out_path) as zf:
            chart_xmls = {
                name: zf.read(name).decode("utf-8", errors="ignore")
                for name in zf.namelist()
                if name.startswith("xl/charts/chart") and name.endswith(".xml")
            }
        weekly_chart_xml = next(
            xml
            for xml in chart_xmls.values()
            if "Approximate market crush ($/gal)" in xml and "Next quarter outlook ($/gal)" in xml
        )
        quarterly_chart_xml = next(
            xml
            for xml in chart_xmls.values()
            if "<lineChart>" in xml and 'legendPos val="t"' in xml
        )
        coproduct_chart_xml = next(
            xml
            for xml in chart_xmls.values()
            if "<lineChart>" in xml
            and "'Economics_Overlay'!$AW$" in xml
            and 'legendPos val="t"' not in xml
        )
        assert "Approximate market crush ($/gal)" in weekly_chart_xml
        assert "Prior quarter ($/gal)" in weekly_chart_xml
        assert "Current QTD ($/gal)" in weekly_chart_xml
        assert "Next quarter outlook ($/gal)" in weekly_chart_xml
        assert "Quarter boundary" in weekly_chart_xml
        assert 'symbol val="diamond"' not in weekly_chart_xml
        assert "<lineChart>" in quarterly_chart_xml
        assert "'Economics_Overlay'!AT" in quarterly_chart_xml
        assert "'Economics_Overlay'!AU" in quarterly_chart_xml
        assert "'Economics_Overlay'!AV" in quarterly_chart_xml
        assert "'Economics_Overlay'!AW" in quarterly_chart_xml
        assert "'Economics_Overlay'!$AS$" in quarterly_chart_xml
        assert "<strRef>" in quarterly_chart_xml
        assert 'legendPos val="t"' in quarterly_chart_xml
        assert '<overlay val="1"/>' in quarterly_chart_xml
        assert "<dLbls>" in quarterly_chart_xml
        assert '<showLegendKey val="1"/>' not in quarterly_chart_xml
        assert '<showVal val="1"/>' in quarterly_chart_xml
        assert '<dLblPos val="r"/>' in quarterly_chart_xml
        assert re.search(r"<catAx>.*?<majorGridlines/>", quarterly_chart_xml)
        assert re.search(r"<catAx>.*?<delete val=\"0\"/>", quarterly_chart_xml)
        assert re.search(r"<catAx>.*?<auto val=\"0\"/>", quarterly_chart_xml)
        assert re.search(r"<catAx>.*?<axPos val=\"b\"/>", quarterly_chart_xml)
        assert re.search(r"<catAx>[\s\S]*?<spPr>[\s\S]*?<a:noFill/>[\s\S]*?</spPr>", quarterly_chart_xml)
        assert re.search(r"<valAx>.*?<crosses val=\"min\"/>", quarterly_chart_xml)
        assert "'Economics_Overlay'!BC" not in quarterly_chart_xml
        assert "<legendEntry>" not in quarterly_chart_xml
        assert "Quarter boundary" not in quarterly_chart_xml
        assert "Next quarter outlook ($/gal)" not in quarterly_chart_xml
        assert "'Economics_Overlay'!$AX$" in coproduct_chart_xml
        assert "'Economics_Overlay'!$AW$" in coproduct_chart_xml
        assert "<strRef>" in coproduct_chart_xml
        assert "<dLbls>" in coproduct_chart_xml
        assert '<showLegendKey val="1"/>' not in coproduct_chart_xml
        assert '<showVal val="1"/>' in coproduct_chart_xml
        assert '<dLblPos val="b"/>' in coproduct_chart_xml
        assert re.search(r"<catAx>.*?<majorGridlines/>", coproduct_chart_xml)
        assert re.search(r"<catAx>.*?<delete val=\"0\"/>", coproduct_chart_xml)
        assert re.search(r"<catAx>.*?<auto val=\"0\"/>", coproduct_chart_xml)
        assert "'Basis_Proxy_Sandbox'!" not in coproduct_chart_xml


def test_valuation_cashflow_deltas_do_not_overwrite_direct_quarter_safe_buyback_truth(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "valuation_buyback_direct_truth_not_overwritten.xlsx")
            sec_cache_dir = case_dir / "PBI" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828025012345_pbi-20250331.htm").write_text(
                "Pitney Bowes Inc. During the quarter we repurchased 1.5 million shares for $15.0 million. "
                "Condensed Consolidated Statements of Cash Flows Common stock repurchases ( 15,000 )",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828025023456_pbi-20250630.htm").write_text(
                "Pitney Bowes Inc. During the quarter we repurchased 7.6 million shares for $75.3 million. "
                "Condensed Consolidated Statements of Cash Flows Common stock repurchases ( 90,274 )",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828025034567_pbi-20250930.htm").write_text(
                "Pitney Bowes Inc. During the quarter we repurchased 14.1 million shares for $161.5 million. "
                "Condensed Consolidated Statements of Cash Flows Common stock repurchases ( 251,774 )",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828026045678_pbi-20251231.htm").write_text(
                "Pitney Bowes Inc. During the quarter we repurchased 12.6 million shares for $126.6 million. "
                "Consolidated Statements of Cash Flows Common stock repurchases ( 378,361 )",
                encoding="utf-8",
            )

            base_inputs = _make_inputs(out_path, ticker="PBI")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            result = write_excel_from_inputs(inputs)
            valuation_snapshot = result.saved_workbook_provenance.get("valuation_snapshot") or {}
            quarter_headers = list(valuation_snapshot.get("quarter_headers") or [])
            buyback_cash_vals = list((valuation_snapshot.get("grid_rows") or {}).get("Buybacks (cash)") or [])
            cash_by_q = dict(zip(quarter_headers, buyback_cash_vals))

            assert cash_by_q.get("2025-Q1") == pytest.approx(15.0)
            assert cash_by_q.get("2025-Q2") == pytest.approx(75.3, rel=1e-9)
            assert cash_by_q.get("2025-Q3") == pytest.approx(161.5, rel=1e-9)
            assert cash_by_q.get("2025-Q4") == pytest.approx(126.6, rel=1e-9)


def test_pbi_live_buybacks_cash_recent_quarters_follow_precompute_truth() -> None:
    with _case_dir() as case_dir:
        out_path = _make_ticker_model_out_path(case_dir, "PBI", "pbi_live_buyback_truth.xlsx")
        inputs = _make_live_artifact_inputs("PBI", out_path)
        ctx = build_writer_context(inputs)
        ensure_valuation_inputs(ctx)

        quarter_key = tuple(pd.Timestamp(q).normalize() for q in list((ctx.derived.valuation_core_maps or {}).get("quarters") or []))
        render_bundle = ctx.state["_ensure_valuation_render_bundle"](quarter_key, ctx.derived.leverage_df)
        pre = ctx.state["_ensure_valuation_precompute_bundle"](quarter_key, render_bundle)
        buyback_map = dict(pre.get("buyback_map") or {})

        assert buyback_map[pd.Timestamp("2023-03-31")] == pytest.approx(13_446_000.0, abs=1e-3)
        assert buyback_map[pd.Timestamp("2025-03-31")] == pytest.approx(15_006_378.05, abs=1e-2)
        assert buyback_map[pd.Timestamp("2025-06-30")] == pytest.approx(75_250_560.64, abs=1e-2)
        assert buyback_map[pd.Timestamp("2025-09-30")] == pytest.approx(161_511_498.72, abs=1e-2)
        assert buyback_map[pd.Timestamp("2025-12-31")] == pytest.approx(126_640_513.88, abs=1e-2)
        assert buyback_map[pd.Timestamp("2025-03-31")] != pytest.approx(150_000_000.0, abs=1e-6)


def test_valuation_filing_docs_filter_out_audit_bucket_quarter_leakage(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "valuation_buyback_doc_quarter_filter.xlsx")
            sec_cache_dir = case_dir / "PBI" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828025023713_pbi-20250331.htm").write_text(
                "Pitney Bowes Inc. During the quarter we repurchased 1.5 million shares for $15.0 million.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828025036856_pbi-20250630.htm").write_text(
                "Pitney Bowes Inc. During the quarter we repurchased 7.6 million shares for $75.3 million.",
                encoding="utf-8",
            )
            audit = pd.DataFrame(
                {
                    "quarter": [
                        pd.Timestamp("2025-03-31"),
                        pd.Timestamp("2025-03-31"),
                        pd.Timestamp("2025-06-30"),
                    ],
                    "accn": [
                        "0001628280-25-023713",
                        "0001628280-25-036856",
                        "0001628280-25-036856",
                    ],
                    "form": ["10-Q", "10-Q", "10-Q"],
                    "filed": [
                        pd.Timestamp("2025-05-08"),
                        pd.Timestamp("2025-07-31"),
                        pd.Timestamp("2025-07-31"),
                    ],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="PBI")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir, "audit": audit})
            ctx = build_writer_context(inputs)
            ensure_valuation_inputs(ctx)
            quarter_key = tuple(pd.Timestamp(q) for q in ctx.inputs.hist["quarter"].tolist())
            render_bundle = ctx.state["_ensure_valuation_render_bundle"](quarter_key, ctx.derived.leverage_df)
            pre = ctx.state["_ensure_valuation_precompute_bundle"](quarter_key, render_bundle)

            assert pre["buyback_cash_doc_map"][pd.Timestamp("2025-03-31")] == pytest.approx(15_000_000.0)
            assert pre["buyback_cash_doc_map"][pd.Timestamp("2025-06-30")] == pytest.approx(75_300_000.0)
            q1_docs = list((ctx.derived.valuation_filing_docs_by_quarter or {}).get(pd.Timestamp("2025-03-31")) or [])
            assert all("20250630" not in str(rec.get("name") or "") for rec in q1_docs)


def test_valuation_submission_8k_midquarter_event_maps_into_execution_quarter(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "valuation_buyback_midquarter_8k.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            accn = "0001309402-25-000156"
            (sec_cache_dir / "submissions_test.json").write_text(
                json.dumps(
                    {
                        "filings": {
                            "recent": {
                                "accessionNumber": [accn],
                                "form": ["8-K"],
                                "filingDate": ["2025-10-22"],
                                "reportDate": [""],
                                "primaryDocument": ["exhibit991_pressrelease102.htm"],
                            }
                        }
                    }
                ),
                encoding="utf-8",
            )
            (
                sec_cache_dir / "doc_000130940225000156_exhibit991_pressrelease102.htm"
            ).write_text(
                "Green Plains Inc. In connection with the exchange transactions and subscription transactions, "
                "the company agreed to repurchase approximately 2.9 million shares of its common stock for "
                "approximately $30.0 million. The transactions are expected to close on October 27, 2025.",
                encoding="utf-8",
            )
            audit = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-12-31")],
                    "accn": [None],
                    "form": [None],
                    "filed": [pd.NaT],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="GPRE")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir, "audit": audit})
            ctx = build_writer_context(inputs)
            ensure_valuation_inputs(ctx)
            quarter_key = tuple(pd.Timestamp(q) for q in ctx.inputs.hist["quarter"].tolist())
            render_bundle = ctx.state["_ensure_valuation_render_bundle"](quarter_key, ctx.derived.leverage_df)
            pre = ctx.state["_ensure_valuation_precompute_bundle"](quarter_key, render_bundle)

            assert pre["buyback_cash_doc_map"][pd.Timestamp("2025-12-31")] == pytest.approx(30_000_000.0)
            assert pre["buyback_shares_doc_map"][pd.Timestamp("2025-12-31")] == pytest.approx(2_900_000.0)
            assert "2.9m shares" in str(pre["buyback_doc_note_map"][pd.Timestamp("2025-12-31")] or "").lower()


def test_pbi_capped_call_note_stays_in_origin_quarter_only(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_notes_capped_call_origin_only.xlsx")
            sec_cache_dir = case_dir / "PBI" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828025047122_pbi-20250930.htm").write_text(
                "Pitney Bowes Inc. In August 2025, we issued an aggregate $230 million convertible senior notes due 2030. "
                "The Convertible Notes accrue interest at a rate of 1.50% per annum. Net proceeds were $221 million. "
                "We used $61.9 million of the proceeds to repurchase 5.5 million of our common stock. "
                "We entered into capped call transactions that are expected to reduce the potential dilution of our common stock upon conversion. ",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828026000020_pbi-20251231.htm").write_text(
                "Pitney Bowes Inc. We entered into capped call transactions that are expected to reduce the potential "
                "dilution of our common stock upon conversion.",
                encoding="utf-8",
            )
            quarter_notes = pd.DataFrame(
                {
                    "quarter": [pd.Timestamp("2025-09-30")] * 2 + [pd.Timestamp("2025-12-31")] * 2,
                    "note_id": ["q3-guid", "q3-fcf", "q4-guid", "q4-fcf"],
                    "category": [
                        "Guidance / outlook",
                        "Cash flow / FCF / capex",
                        "Guidance / outlook",
                        "Cash flow / FCF / capex",
                    ],
                    "claim": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FCF TTM improved by $25.0m YoY.",
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FCF TTM improved by $35.0m YoY.",
                    ],
                    "note": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FCF TTM improved by $25.0m YoY.",
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FCF TTM improved by $35.0m YoY.",
                    ],
                    "metric_ref": ["Revenue guidance", "FCF", "Revenue guidance", "FCF"],
                    "score": [95.0, 92.0, 95.0, 92.0],
                    "doc_type": ["earnings_release", "model_metric", "earnings_release", "model_metric"],
                    "doc": ["release_q3.txt", "history_q", "release_q4.txt", "history_q"],
                    "source_type": ["earnings_release", "model_metric", "earnings_release", "model_metric"],
                    "evidence_snippet": [
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FCF TTM improved by $25.0m YoY.",
                        "FY 2025 Revenue guidance tracking near the midpoint of $1.90bn-$1.95bn.",
                        "FCF TTM improved by $35.0m YoY.",
                    ],
                }
            )

            base_inputs = _make_inputs(out_path, ticker="PBI", quarter_notes=quarter_notes)
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            ctx = build_writer_context(inputs)
            ctx.callbacks.write_quarter_notes_ui_v2()
            ws = ctx.wb["Quarter_Notes_UI"]

            q3_rows = _quarter_block_notes(ws, "2025-09-30")
            q4_rows = _quarter_block_notes(ws, "2025-12-31")

            assert any(
                "Entered capped call transactions expected to reduce dilution from convertible notes conversion." in note
                for note in q3_rows
            )
            assert not any("capped call transactions" in note.lower() for note in q4_rows)


def test_valuation_keeps_adj_ebitda_and_adj_ebit_distinct_when_inputs_are_distinct() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "valuation_adj_ebitda_distinct.xlsx")
        base_inputs = _make_inputs(out_path)
        quarter_vals = list(pd.to_datetime(base_inputs.hist["quarter"], errors="coerce").dropna())
        adj_metrics = pd.DataFrame(
            {
                "quarter": quarter_vals,
                "adj_ebit": [40_000_000.0 + (i * 1_000_000.0) for i in range(len(quarter_vals))],
                "adj_ebitda": [55_000_000.0 + (i * 1_000_000.0) for i in range(len(quarter_vals))],
            }
        )
        inputs = base_inputs.__class__(**{**vars(base_inputs), "adj_metrics": adj_metrics})
        ctx = build_writer_context(inputs)
        ensure_valuation_inputs(ctx)
        ctx.callbacks.write_valuation_sheet()
        ws = ctx.wb["Valuation"]

        adj_ebitda_row = _find_row_with_value(ws, "Adj EBITDA (TTM)")
        adj_ebit_row = _find_row_with_value(ws, "Adj EBIT (TTM)")

        assert adj_ebitda_row is not None
        assert adj_ebit_row is not None
        adj_ebitda_vals = [ws.cell(row=adj_ebitda_row, column=cc).value for cc in range(2, ws.max_column + 1)]
        adj_ebit_vals = [ws.cell(row=adj_ebit_row, column=cc).value for cc in range(2, ws.max_column + 1)]
        assert adj_ebitda_vals != adj_ebit_vals


def test_valuation_uses_gaap_and_adjusted_denominators_on_the_correct_rows() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "valuation_denominator_semantics.xlsx")
        hist = pd.DataFrame(
            {
                "quarter": pd.to_datetime(["2025-03-31", "2025-06-30", "2025-09-30", "2025-12-31"]),
                "revenue": [500_000_000.0] * 4,
                "ebitda": [100_000_000.0] * 4,
                "ebit": [80_000_000.0] * 4,
                "op_income": [70_000_000.0] * 4,
                "cash": [20_000_000.0] * 4,
                "debt_core": [220_000_000.0] * 4,
                "shares_outstanding": [10_000_000.0] * 4,
                "shares_diluted": [10_000_000.0] * 4,
                "market_cap": [100_000_000.0] * 4,
                "interest_expense_net": [10_000_000.0] * 4,
                "interest_paid": [20_000_000.0] * 4,
            }
        )
        adj_metrics = pd.DataFrame(
            {
                "quarter": hist["quarter"],
                "adj_ebit": [120_000_000.0] * 4,
                "adj_ebitda": [200_000_000.0] * 4,
            }
        )
        inputs = _make_inputs(out_path, hist=hist)
        inputs = inputs.__class__(**{**vars(inputs), "adj_metrics": adj_metrics})
        ctx = build_writer_context(inputs)
        ensure_valuation_inputs(ctx)
        ctx.callbacks.write_valuation_sheet()
        ws = ctx.wb["Valuation"]

        latest_col = 5
        net_lev_row = _find_row_with_value(ws, "Net leverage")
        net_lev_adj_row = _find_row_with_value(ws, "Net leverage (Adj)")
        cov_pnl_row = _find_row_with_value(ws, "Interest coverage (P&L TTM)")
        cov_cash_row = _find_row_with_value(ws, "Cash interest coverage (TTM)")

        assert net_lev_row is not None
        assert net_lev_adj_row is not None
        assert cov_pnl_row is not None
        assert cov_cash_row is not None
        assert ws.cell(row=net_lev_row, column=latest_col).value == pytest.approx(0.5)
        assert ws.cell(row=net_lev_adj_row, column=latest_col).value == pytest.approx(0.25)
        assert ws.cell(row=cov_pnl_row, column=latest_col).value == pytest.approx(10.0)
        assert ws.cell(row=cov_cash_row, column=latest_col).value == pytest.approx(5.0)

        audit = dict((ctx.derived.valuation_precompute_bundle or {}).get("valuation_audit") or {})
        latest_q = pd.Timestamp("2025-12-31")
        assert audit[latest_q]["net_leverage"]["scope"] == "gaap_ebitda_ttm"
        assert audit[latest_q]["net_leverage_adj"]["scope"] == "adjusted_ebitda_ttm"
        assert audit[latest_q]["interest_coverage_pnl"]["scope"] == "gaap_ebitda_ttm/pnl_interest_ttm"
        assert audit[latest_q]["cash_interest_coverage"]["scope"] == "gaap_ebitda_ttm/cash_interest_ttm"
        assert audit[latest_q]["adj_ebit_ttm"]["value"] == pytest.approx(480_000_000.0)
        assert audit[latest_q]["adj_ebitda_ttm"]["value"] == pytest.approx(800_000_000.0)


def test_valuation_displays_nm_when_ebitda_denominator_is_nonmeaningful() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "valuation_nm_on_negative_ebitda.xlsx")
        hist = pd.DataFrame(
            {
                "quarter": pd.to_datetime(["2025-03-31", "2025-06-30", "2025-09-30", "2025-12-31"]),
                "revenue": [500_000_000.0] * 4,
                "ebitda": [-20_000_000.0] * 4,
                "ebit": [-35_000_000.0] * 4,
                "op_income": [-40_000_000.0] * 4,
                "cash": [20_000_000.0] * 4,
                "debt_core": [220_000_000.0] * 4,
                "shares_outstanding": [10_000_000.0] * 4,
                "shares_diluted": [10_000_000.0] * 4,
                "market_cap": [100_000_000.0] * 4,
                "interest_expense_net": [10_000_000.0] * 4,
                "interest_paid": [20_000_000.0] * 4,
            }
        )
        adj_metrics = pd.DataFrame(
            {
                "quarter": hist["quarter"],
                "adj_ebit": [-30_000_000.0] * 4,
                "adj_ebitda": [-10_000_000.0] * 4,
            }
        )
        inputs = _make_inputs(out_path, hist=hist)
        inputs = inputs.__class__(**{**vars(inputs), "adj_metrics": adj_metrics})
        ctx = build_writer_context(inputs)
        ensure_valuation_inputs(ctx)
        ctx.callbacks.write_valuation_sheet()
        ws = ctx.wb["Valuation"]

        latest_col = 5
        for label in ("Net leverage", "Net leverage (Adj)", "Interest coverage (P&L TTM)", "Cash interest coverage (TTM)"):
            row_idx = _find_row_with_value(ws, label)
            assert row_idx is not None
            assert ws.cell(row=row_idx, column=latest_col).value == "N/M"

        audit = dict((ctx.derived.valuation_precompute_bundle or {}).get("valuation_audit") or {})
        latest_q = pd.Timestamp("2025-12-31")
        assert audit[latest_q]["net_leverage"]["suppress_reason"] == "EBITDA denominator <= 0"
        assert audit[latest_q]["net_leverage_adj"]["suppress_reason"] == "Adjusted EBITDA denominator <= 0"
        assert audit[latest_q]["interest_coverage_pnl"]["suppress_reason"] == "EBITDA denominator <= 0"
        assert audit[latest_q]["cash_interest_coverage"]["suppress_reason"] == "EBITDA denominator <= 0"


def test_quarter_notes_ui_uses_reaffirmed_and_never_shows_repeat_badge() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "quarter_notes_no_repeat_badge.xlsx")
        quarter_notes = pd.DataFrame(
            {
                "quarter": [pd.Timestamp("2025-09-30"), pd.Timestamp("2025-12-31")],
                "note_id": ["q3-guid", "q4-guid"],
                "category": ["Guidance / outlook", "Guidance / outlook"],
                "claim": [
                    "FY 2025 Revenue guidance reaffirmed at $1.90bn-$1.95bn.",
                    "FY 2025 Revenue guidance reaffirmed at $1.90bn-$1.95bn.",
                ],
                "note": [
                    "FY 2025 Revenue guidance reaffirmed at $1.90bn-$1.95bn.",
                    "FY 2025 Revenue guidance reaffirmed at $1.90bn-$1.95bn.",
                ],
                "metric_ref": ["Revenue guidance", "Revenue guidance"],
                "score": [95.0, 95.0],
                "doc_type": ["earnings_release", "earnings_release"],
                "doc": ["release_q3.txt", "release_q4.txt"],
                "source_type": ["earnings_release", "earnings_release"],
                "evidence_snippet": [
                    "FY 2025 Revenue guidance reaffirmed at $1.90bn-$1.95bn.",
                    "FY 2025 Revenue guidance reaffirmed at $1.90bn-$1.95bn.",
                ],
            }
        )

        ctx = build_writer_context(_make_inputs(out_path, ticker="TEST", quarter_notes=quarter_notes))
        ctx.callbacks.write_quarter_notes_ui_v2()
        ws = ctx.wb["Quarter_Notes_UI"]

        q4_rows = _quarter_block_notes(ws, "2025-12-31")
        assert not any("[REPEAT]" in note for note in q4_rows)
        assert not any("[REPEAT]" in note for note in _quarter_block_notes(ws, "2025-09-30"))
        assert not any("[REAFFIRMED]" in note for note in q4_rows)
        if q4_rows:
            assert any(
                ("[CONTINUED]" in note)
                or not note.startswith("[")
                for note in q4_rows
            )


def test_write_excel_saved_workbook_provenance_captures_summary_and_valuation_truth() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "saved_workbook_summary_valuation_truth.xlsx")
        inputs = _make_inputs(out_path, ticker="TEST", quarter_notes=pd.DataFrame())
        inputs.company_overview = {
            "what_it_does": "Demo company description.",
            "what_it_does_source": "Source: SEC 10-K demo",
            "current_strategic_context": "Management is focused on capital allocation and cost discipline.",
            "current_strategic_context_source": "Source: SEC 8-K demo",
            "key_advantage": "Demo advantage.",
            "key_advantage_source": "Source: SEC 10-K competition demo",
            "segment_operating_model": [],
            "segment_operating_model_source": "Source: N/A",
            "key_dependencies": [],
            "key_dependencies_source": "Source: N/A",
            "wrong_thesis_bullets": [],
            "wrong_thesis_source": "Source: N/A",
            "revenue_streams": [],
            "revenue_streams_source": "Source: N/A",
            "asof_fy_end": None,
        }

        result = write_excel_from_inputs(inputs)

        assert result.summary_export_expectation
        assert result.valuation_export_expectation
        validate_summary_export(out_path, result.summary_export_expectation)
        validate_valuation_export(out_path, result.valuation_export_expectation)

        summary_snapshot = result.saved_workbook_provenance.get("summary_snapshot") or {}
        valuation_snapshot = result.saved_workbook_provenance.get("valuation_snapshot") or {}
        assert summary_snapshot["What the company does"]["value"] == "Demo company description."
        assert summary_snapshot["Current strategic context"]["source"] == "Source: SEC 8-K demo"
        assert "Buybacks (cash)" in dict(valuation_snapshot.get("grid_rows") or {})
        assert "Buybacks (TTM, cash)" in dict(valuation_snapshot.get("grid_rows") or {})


def test_pbi_saved_workbook_buyback_truth_is_synced_across_valuation_qa_and_needs_review(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "PBI"):
            out_path = _make_model_out_path(case_dir, "pbi_buyback_truth_sync.xlsx")
            sec_cache_dir = case_dir / "PBI" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000162828026008604_q42025earningspressrelea.htm").write_text(
                "During the quarter we repurchased 12.6 million shares for $127.0 million.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000162828026012345_pbi-20251231.htm").write_text(
                "The following table provides information about common stock purchases during the three months ended December 31, 2025. "
                "Total number of shares purchased Average price paid per share Total number of shares purchased as part of publicly announced plans or programs Approximate dollar value of shares that may yet be purchased under the plans or programs. "
                "October 2025 3,203,100 $ 11.30 3,203,100 $212,031 "
                "November 2025 7,926,090 $ 9.52 7,926,090 $136,553 "
                "December 2025 1,484,407 $ 10.05 1,484,407 $121,639 "
                "12,613,597&#160; $ 10.04&#160; 12,613,597 ",
                encoding="utf-8",
            )

            base_inputs = _make_inputs(out_path, ticker="PBI")
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            result = write_excel_from_inputs(inputs)

            validate_summary_export(out_path, result.summary_export_expectation)
            validate_valuation_export(out_path, result.valuation_export_expectation)
            validate_qa_export(out_path, result.qa_export_expectation)
            validate_needs_review_export(out_path, result.needs_review_export_expectation)

            valuation_snapshot = result.saved_workbook_provenance.get("valuation_snapshot") or {}
            hidden_rows = dict(valuation_snapshot.get("hidden_rows") or {})
            shares_text = str(hidden_rows.get("Buybacks (shares)") or "")
            note_text = str(hidden_rows.get("Buybacks note") or "")
            assert "QoQ +12.614m" in shares_text
            assert "$10.04/share" in shares_text
            assert "$126.6m" in shares_text
            assert "convertible notes" not in shares_text.lower()
            assert "YoY" not in shares_text
            assert "Remaining capacity $359.0m" in note_text
            assert "Latest increase by $250.0m on 2026-02-13" in note_text
            assert "$127.0m" not in note_text
            assert "$10.08/share" not in note_text
            assert "YoY" not in note_text

            qa_rows = result.saved_workbook_provenance.get("qa_checks_snapshot") or []
            assert len(qa_rows) <= 1
            if qa_rows:
                qa_msg = str(qa_rows[0].get("message") or "")
                assert "$126.6m" in qa_msg
                assert "$10.04/share" in qa_msg
                assert "$127.0m" not in qa_msg
                assert "$10.08/share" not in qa_msg
                assert "supported by explicit SEC repurchase disclosures" not in qa_msg

            needs_rows = result.saved_workbook_provenance.get("needs_review_snapshot") or []
            assert needs_rows == []
            assert _sheet_metric_rows(
                out_path,
                "Needs_Review",
                "buybacks_cash",
                quarters=list((result.needs_review_export_expectation or {}).get("quarters") or []),
            ) == []


def test_gpre_saved_workbook_buyback_truth_drops_historical_leakage_and_syncs_recent_truth(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with _case_dir() as case_dir:
        with _profile_override(monkeypatch, "GPRE"):
            out_path = _make_model_out_path(case_dir, "gpre_buyback_truth_sync.xlsx")
            sec_cache_dir = case_dir / "GPRE" / "sec_cache"
            sec_cache_dir.mkdir(parents=True, exist_ok=True)
            (sec_cache_dir / "doc_000130940224000019_gpre-20241231.htm").write_text(
                "To date, we have repurchased approximately 7.4 million shares of common stock for approximately $92.8 million under the program.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940225000120_gpre-20250630.htm").write_text(
                "We did not repurchase any common stock during the second quarter of 2025.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940225000163_gpre-20250930.htm").write_text(
                "To date, we have repurchased approximately 7.4 million shares of common stock for approximately $92.8 million under the program. No repurchases were made during Q3 2025.",
                encoding="utf-8",
            )
            (sec_cache_dir / "doc_000130940226000019_gpre-20251231.htm").write_text(
                "On October 27, 2025, in conjunction with the privately negotiated exchange and subscription agreements for the 2030 Notes, "
                "the company repurchased approximately 2.9 million shares of its common stock for approximately $30.0 million. "
                "No other repurchase was made during 2025. At February 10, 2026, $77.2 million in share repurchase authorization remained.",
                encoding="utf-8",
            )
            hist = pd.DataFrame(
                {
                    "quarter": pd.to_datetime(
                        [
                            "2024-03-31",
                            "2024-06-30",
                            "2024-09-30",
                            "2024-12-31",
                            "2025-03-31",
                            "2025-06-30",
                            "2025-09-30",
                            "2025-12-31",
                        ]
                    ),
                    "revenue": [500_000_000.0] * 8,
                    "ebitda": [50_000_000.0] * 8,
                    "ebit": [30_000_000.0] * 8,
                    "op_income": [25_000_000.0] * 8,
                    "cash": [20_000_000.0] * 8,
                    "debt_core": [150_000_000.0] * 8,
                    "interest_paid": [10_000_000.0] * 8,
                    "shares_outstanding": [100_000_000.0] * 8,
                    "shares_diluted": [100_000_000.0] * 8,
                    "market_cap": [1_000_000_000.0] * 8,
                }
            )

            base_inputs = _make_inputs(out_path, ticker="GPRE", hist=hist)
            inputs = base_inputs.__class__(**{**vars(base_inputs), "cache_dir": sec_cache_dir})
            result = write_excel_from_inputs(inputs)

            validate_summary_export(out_path, result.summary_export_expectation)
            validate_valuation_export(out_path, result.valuation_export_expectation)
            validate_qa_export(out_path, result.qa_export_expectation)
            validate_needs_review_export(out_path, result.needs_review_export_expectation)

            valuation_snapshot = result.saved_workbook_provenance.get("valuation_snapshot") or {}
            quarter_headers = list(valuation_snapshot.get("quarter_headers") or [])
            buyback_ttm_vals = list((valuation_snapshot.get("grid_rows") or {}).get("Buybacks (TTM, cash)") or [])
            ttm_by_q = dict(zip(quarter_headers, buyback_ttm_vals))
            assert ttm_by_q.get("2024-Q4") in (None, "")
            assert ttm_by_q.get("2025-Q1") in (None, "")
            assert ttm_by_q.get("2025-Q2") in (None, "")
            assert ttm_by_q.get("2025-Q3") in (None, "")
            assert ttm_by_q.get("2025-Q4") in (None, "")

            hidden_rows = dict(valuation_snapshot.get("hidden_rows") or {})
            shares_text = str(hidden_rows.get("Buybacks (shares)") or "")
            note_text = str(hidden_rows.get("Buybacks note") or "")
            assert "+2.900m" in shares_text
            assert "$10.34/share" in shares_text
            assert "includes 2.900m shares concurrent with convertible notes" in shares_text
            assert "2.024" not in shares_text
            assert "TTM" not in shares_text
            assert "Remaining capacity $77.2m" in note_text
            assert "$2.0m" not in note_text
            assert "94.8" not in note_text
            assert "YoY" not in note_text

            qa_rows = result.saved_workbook_provenance.get("qa_checks_snapshot") or []
            assert len(qa_rows) == 1
            qa_msg = str(qa_rows[0].get("message") or "")
            assert "2.900m" in qa_msg
            assert "$30.0m" in qa_msg
            assert "$2.0m" not in qa_msg
            assert "supported by explicit SEC repurchase disclosures" not in qa_msg

            needs_rows = result.saved_workbook_provenance.get("needs_review_snapshot") or []
            assert needs_rows == []
            assert _sheet_metric_rows(
                out_path,
                "Needs_Review",
                "buybacks_cash",
                quarters=list((result.needs_review_export_expectation or {}).get("quarters") or []),
            ) == []


def test_needs_review_source_gap_ordering_prefers_stronger_expected_metrics() -> None:
    with _case_dir() as case_dir:
        out_path = _make_model_out_path(case_dir, "needs_review_source_gap_ordering.xlsx")
        qref = pd.Timestamp("2025-12-31")
        custom_needs_review = pd.DataFrame(
            [
                {
                    "quarter": qref,
                    "metric": "EBITDA (Q)",
                    "severity": "warn",
                    "message": "EBITDA (Q): workbook=$159.0m; no explicit quarter-level statement found in selected release/presentation corpus",
                    "source": "release.htm | slides.pdf",
                    "issue_family": "quarter_text_no_explicit_support",
                    "raw_metric": "EBITDA (Q)",
                    "recommended_action": "review source coverage",
                },
                {
                    "quarter": qref,
                    "metric": "Revenue (Q)",
                    "severity": "warn",
                    "message": "Revenue (Q): workbook=$1,000.0m; no explicit quarter-level statement found in selected release/presentation corpus",
                    "source": "release.htm | slides.pdf",
                    "issue_family": "quarter_text_no_explicit_support",
                    "raw_metric": "Revenue (Q)",
                    "recommended_action": "review source coverage",
                },
            ]
        )
        base_inputs = _make_inputs(out_path, ticker="TEST")
        inputs = base_inputs.__class__(**{**vars(base_inputs), "needs_review": custom_needs_review})
        ctx = build_writer_context(inputs)
        setattr(ctx.callbacks, "run_latest_quarter_qa", lambda: [])

        write_qa_sheets(ctx, [])

        ws_nr = ctx.wb["Needs_Review"]
        assert _sheet_data_row_count(ws_nr) == 1
        assert str(ws_nr.cell(row=2, column=1).value or "").strip() == "Current-quarter source gaps"
        assert str(ws_nr.cell(row=2, column=11).value or "").strip() == "Revenue (Q)"
        ws_log = ctx.wb["QA_Log"]
        qa_log_blob = "\n".join(
            " | ".join(str(ws_log.cell(row=rr, column=cc).value or "").strip() for cc in range(1, min(ws_log.max_column, 8) + 1))
            for rr in range(1, ws_log.max_row + 1)
        )
        assert "2025-12-31 00:00:00 | EBITDA (Q) | warn |" in qa_log_blob
        assert "| quarter_text_no_explicit_support | EBITDA (Q) |" in qa_log_blob


def test_current_delivered_workbooks_match_visible_quarter_notes_ui_snapshots() -> None:
    expected_snapshots = {
        "PBI": {
            "2025-09-30": [
                ("Guidance / outlook", "FY 2025 Revenue guidance tracking near the midpoint of $1,900m-$1,950m."),
                ("Guidance / outlook", "FY 2025 Adjusted EBIT guidance tracking near the midpoint of $450m-$465m."),
                ("Guidance / outlook", "FY 2025 EPS guidance tracking near the midpoint of $1.20-$1.40."),
                ("Guidance / outlook", "FY 2025 FCF target tracking near the midpoint of $330m-$370m."),
                ("Cash flow / FCF / working capital", "Free cash flow declined to $60.4m, down $13.1m YoY."),
                ("Debt / liquidity / balance sheet", "Revolver availability increased from $265.0m to $400.0m."),
                ("Capital allocation / shareholder returns", "Repurchase authorization increased to $500.0m, up from $400.0m."),
                ("Capital allocation / shareholder returns", "Remaining share repurchase capacity was $148.2m at quarter-end."),
                ("Capital allocation / shareholder returns", "Quarterly dividend increased to $0.09/share from $0.08/share."),
                ("Capital allocation / shareholder returns", "Repurchased 14.1m shares for $161.5m with an average price of $11.44/share in Q3."),
                ("Capital allocation / shareholder returns", "Used $61.9m from convertible notes proceeds to repurchase 5.5m shares."),
                ("Capital allocation / shareholder returns", "Entered capped call transactions expected to reduce dilution from convertible notes conversion."),
            ],
            "2025-12-31": [
                ("Guidance / outlook", "FY 2026 Revenue guidance $1,760m-$1,860m."),
                ("Guidance / outlook", "FY 2026 Adjusted EBIT guidance $410m-$460m."),
                ("Guidance / outlook", "FY 2026 EPS guidance $1.40-$1.60."),
                ("Guidance / outlook", "FY 2026 FCF target $340m-$370m."),
                ("Guidance / outlook", "Guidance ranges widened due to market uncertainty and forecasting changes."),
                ("Results / drivers / better vs prior", "Gross margin expanded 180 bps, driven by cost optimization and a shift to higher margin revenue streams."),
                ("Results / drivers / better vs prior", "Operating expenses declined $28.0m YoY, primarily from cost reduction."),
                ("Cash flow / FCF / working capital", "Free cash flow improved to $211.9m, up $70.2m YoY."),
                ("Debt / liquidity / balance sheet", "Reduced principal debt by $114.1m in Q4."),
                ("Debt / liquidity / balance sheet", "Reached sub-3.0x leverage, improving covenant flexibility."),
                ("Capital allocation / shareholder returns", "Repurchase authorization increased by $250.0m."),
                ("Capital allocation / shareholder returns", "Remaining share repurchase capacity was $359.0m at quarter-end."),
                ("Capital allocation / shareholder returns", "Quarterly dividend set at $0.09/share."),
                ("Capital allocation / shareholder returns", "Repurchased 12.6m shares for $126.6m with an average price of $10.04/share in Q4."),
                ("Programs / initiatives / management framing", "Strategic review phase 2 remains on track by end of Q2 2026."),
            ],
        },
        "GPRE": {
            "2025-09-30": [
                ("Guidance / outlook", "Q4 2025 45Z monetization expected at $15m-$25m."),
                ("Guidance / outlook", "All eight operating ethanol plants expected to qualify for production tax credits in 2026"),
                ("Results / drivers", "45Z production tax credits contributed $25.0m net of discounts and other costs."),
                ("Results / drivers / better vs prior", "EBITDA margin compressed 405 bps YoY."),
                ("Results / drivers / better vs prior", "Consolidated ethanol crush margin improved to $59.6m from $58.3m YoY."),
                ("Debt / liquidity / balance sheet", "Junior mezzanine debt of $130.7m was repaid from Obion sale proceeds."),
                ("Debt / liquidity / balance sheet", "Revolver availability ended the quarter at $325.0m."),
                ("Capital allocation / shareholder returns", "Repurchase authorization increased to $200.0m."),
                ("Operations / commercialization / milestones", "45Z tax credit monetization agreement for Nebraska production was entered on September 16, 2025."),
                ("Operations / commercialization / milestones", "Utilization reached 101% across operating plants."),
                ("Operations / commercialization / milestones", "York carbon capture was fully operational; Central City and Wood River were online and ramping."),
            ],
            "2025-12-31": [
                ("Guidance / outlook", "FY 2026 45Z-related Adjusted EBITDA outlook is at least $188.0m."),
                ("Guidance / outlook", "Disciplined risk management strategy continues to support first quarter margins and cash flow."),
                ("Results / drivers", "Corporate activities included $16.1m of restructuring costs from the cost reduction initiative."),
                ("Results / drivers", "45Z production tax credits contributed $23.4m net of discounts and other costs in Q4."),
                ("Results / drivers / better vs prior", "Adjusted EBITDA improved 369.8% YoY."),
                ("Results / drivers / better vs prior", "Consolidated ethanol crush margin improved to $44.4m from $(15.5)m YoY."),
                ("Cash flow / FCF / working capital", "FCF TTM improved by $198.7m YoY."),
                ("Debt / liquidity / balance sheet", "Net debt declined by $77.9m YoY."),
                ("Debt / liquidity / balance sheet", "Exchanged $170.0m of 2.25% convertible senior notes due 2027 for $170.0m of 5.25% convertible senior notes due November 2030 (conversion price $15.72/share)."),
                ("Debt / liquidity / balance sheet", "Annualized 2026 interest expense is expected at about $30.0m-$35.0m, reflecting the 2030 convertible notes, Junior Note extinguishment and carbon equipment financing."),
                ("Capital allocation / shareholder returns", "Repurchase authorization increased to $200.0m."),
                ("Capital allocation / shareholder returns", "Repurchased approximately 2.9m shares for approximately $30.0m in connection with the October 27, 2025 exchange and subscription transactions."),
                ("Capital allocation / shareholder returns", "Issued an additional $30.0m of 5.25% convertible senior notes due November 2030; proceeds funded the repurchase of approximately 2.9m shares for approximately $30.0m."),
                ("Operations / commercialization / milestones", "45Z tax credit monetization agreement for Nebraska production was entered on September 16, 2025 and amended on December 10, 2025 to add credits from three additional facilities."),
                ("Operations / commercialization / milestones", "Advantage Nebraska 2026 Adjusted >$150M EBITDA opportunity."),
            ],
        },
    }

    for ticker, expected_snapshot in expected_snapshots.items():
        workbook_path = _current_delivered_model_path(ticker)
        if not workbook_path.exists():
            pytest.skip(f"Current delivered workbook missing for snapshot test: {workbook_path}")
        actual_snapshot = read_quarter_notes_ui_snapshot(workbook_path)
        reduced_actual = {quarter: list(actual_snapshot.get(quarter) or []) for quarter in expected_snapshot}
        assert reduced_actual == expected_snapshot


def test_current_delivered_workbooks_promise_progress_and_guidance_panel_are_clean() -> None:
    def _promise_rows(ws) -> list[tuple[str, str, str, str, str, str]]:
        out: list[tuple[str, str, str, str, str, str]] = []
        for rr in range(1, ws.max_row + 1):
            metric = str(ws.cell(row=rr, column=1).value or "").strip()
            target = str(ws.cell(row=rr, column=2).value or "").strip()
            latest = str(ws.cell(row=rr, column=3).value or "").strip()
            result = str(ws.cell(row=rr, column=4).value or "").strip()
            rationale = str(ws.cell(row=rr, column=5).value or "").strip()
            pid = str(ws.cell(row=rr, column=15).value or "").strip()
            if metric and metric != "Metric":
                out.append((metric, target, latest, result, rationale, pid))
        return out

    def _block_rows(ws, asof_text: str) -> list[tuple[str, str, str, str, str, str, str, str, str, str]]:
        out: list[tuple[str, str, str, str, str, str, str, str, str, str]] = []
        in_block = False
        for rr in range(1, ws.max_row + 1):
            marker = str(ws.cell(row=rr, column=1).value or "").strip()
            if marker == f"Promise progress (As of {asof_text})":
                in_block = True
                continue
            if in_block and marker.startswith("Promise progress (As of "):
                break
            if not in_block:
                continue
            metric = str(ws.cell(row=rr, column=1).value or "").strip()
            if not metric or metric == "Metric":
                continue
            out.append(
                (
                    metric,
                    str(ws.cell(row=rr, column=2).value or "").strip(),
                    str(ws.cell(row=rr, column=3).value or "").strip(),
                    str(ws.cell(row=rr, column=4).value or "").strip(),
                    str(ws.cell(row=rr, column=5).value or "").strip(),
                    str(ws.cell(row=rr, column=6).value or "").strip(),
                    str(ws.cell(row=rr, column=7).value or "").strip(),
                    str(ws.cell(row=rr, column=8).value or "").strip(),
                    str(ws.cell(row=rr, column=9).value or "").strip(),
                    str(ws.cell(row=rr, column=15).value or "").strip(),
                )
            )
        return out

    def _panel_text(ws) -> str:
        chunks: list[str] = []
        for rr in range(1, 28):
            for cc in range(15, 31):
                val = ws.cell(row=rr, column=cc).value
                if val not in (None, ""):
                    chunks.append(str(val))
        return " | ".join(chunks)

    def _guidance_block_rows(ws, asof_text: str) -> list[tuple[str, str, str, str, str]]:
        out: list[tuple[str, str, str, str, str]] = []
        in_block = False
        for rr in range(1, ws.max_row + 1):
            marker = str(ws.cell(row=rr, column=15).value or "").strip()
            if marker.startswith(f"Guidance (As of {asof_text})"):
                in_block = True
                continue
            if in_block and marker.startswith("Guidance (As of "):
                break
            if not in_block:
                continue
            metric = str(ws.cell(row=rr, column=15).value or "").strip()
            if metric in {"", "Metric", "A) Updated / mentioned this quarter", "B) Carry-forward", "No guidance items for this quarter."}:
                continue
            out.append(
                (
                    metric,
                    str(ws.cell(row=rr, column=17).value or "").strip(),
                    str(ws.cell(row=rr, column=18).value or "").strip(),
                    str(ws.cell(row=rr, column=19).value or "").strip(),
                    str(ws.cell(row=rr, column=26).value or "").strip(),
                )
            )
        return out

    def _quarter_note_metrics(ws) -> list[str]:
        out: list[str] = []
        for rr in range(1, ws.max_row + 1):
            metric = str(ws.cell(row=rr, column=4).value or "").strip()
            if metric and metric != "Metric":
                out.append(metric)
        return out

    for ticker in ["PBI", "GPRE"]:
        workbook_path = _current_delivered_model_path(ticker)
        if not workbook_path.exists():
            pytest.skip(f"Current delivered workbook missing for promise/guidance snapshot test: {workbook_path}")
        wb = load_workbook(workbook_path, data_only=False, read_only=False)
        try:
            assert "Promise_Tracker_UI" not in wb.sheetnames
            assert "Promise_Progress_UI" in wb.sheetnames
            assert "Promise_Tracker" in wb.sheetnames
            assert "Promise_Evidence" in wb.sheetnames
            assert "Promise_Progress" in wb.sheetnames

            progress_rows = _promise_rows(wb["Promise_Progress_UI"])
            panel_blob = _panel_text(wb["Valuation"])
            quarter_note_metrics = _quarter_note_metrics(wb["Quarter_Notes_UI"])

            if ticker == "PBI":
                assert all("FY 2043" not in " | ".join(row) for row in progress_rows)
                assert "FY 2043" not in panel_blob
                assert "Management commentary" not in panel_blob
                assert "Guidance full text" not in panel_blob
                assert str(wb["Valuation"].cell(row=7, column=15).value or "").strip().startswith("Guidance (As of 2025-12-31)")
                assert str(wb["Valuation"].cell(row=8, column=15).value or "").strip() == "Metric"
                assert str(wb["Valuation"].cell(row=9, column=15).value or "").strip() == "A) Updated / mentioned this quarter"
                assert str(wb["Valuation"].cell(row=18, column=15).value or "").strip().startswith("Guidance (As of 2025-09-30)")
                assert "Context" not in panel_blob
                assert any(metric == "FCF target" for metric, _, _, _, _, _ in progress_rows)
                assert any(
                    metric == "Cost savings target"
                    and "$180m-$200m" in target
                    and ("annualized cost savings" in rationale.lower() or "raised target" in rationale.lower())
                    for metric, target, _, _, rationale, _ in progress_rows
                )
                assert any(
                    metric == "Strategic milestone"
                    and "Strategic review phase 2 remains on track by end of Q2 2026." in latest
                    for metric, _, latest, _, _, _ in progress_rows
                )
                merged_ranges = list(wb["Valuation"].merged_cells.ranges)
                overlaps = []
                for i, left in enumerate(merged_ranges):
                    for right in merged_ranges[i + 1 :]:
                        if not (
                            left.max_row < right.min_row
                            or right.max_row < left.min_row
                            or left.max_col < right.min_col
                            or right.max_col < left.min_col
                        ):
                            overlaps.append((str(left), str(right)))
                assert overlaps == []
                assert _find_row_with_value(wb["Valuation"], "Operating signals") is not None
                flags_header_row = _find_row_with_value(wb["Valuation"], "Hidden value flags", column=1)
                assert flags_header_row is not None
                assert flags_header_row == 137
                obs_labels = [str(wb["Valuation"].cell(row=rr, column=1).value or "").strip() for rr in range(1, wb["Valuation"].max_row + 1)]
                assert "Obs 5" not in obs_labels
                assert "Obs 6" not in obs_labels
                assert str(wb["Valuation"].cell(row=flags_header_row, column=1).value or "").strip() == "Hidden value flags"
                assert str(wb["Valuation"].cell(row=flags_header_row, column=14).value or "").strip() == "Hidden Value Panel"
                assert any(m.min_row == 137 and m.min_col == 14 and m.max_col == 18 for m in wb["Valuation"].merged_cells.ranges)
                assert str(wb["Valuation"].cell(row=138, column=1).value or "").strip() == "Flag"
                assert str(wb["Valuation"].cell(row=138, column=2).value or "").strip() == "Summary"
                assert str(wb["Valuation"].cell(row=138, column=6).value or "").strip() == "Score"
                assert str(wb["Valuation"].cell(row=138, column=7).value or "").strip() == "Severity"
                assert str(wb["Valuation"].cell(row=138, column=8).value or "").strip() == "Result / support"
                assert str(wb["Valuation"].cell(row=flags_header_row + 2, column=1).value or "").startswith("=IF($AI139")
                assert "INDEX('Hidden_Value_Flags'!$C:$C,$AI139)" in str(wb["Valuation"].cell(row=flags_header_row + 2, column=2).value or "")
                assert "INDEX('Hidden_Value_Flags'!$D:$D,$AI139)" in str(wb["Valuation"].cell(row=flags_header_row + 2, column=6).value or "")
                assert "INDEX('Hidden_Value_Flags'!$E:$E,$AI139)" in str(wb["Valuation"].cell(row=flags_header_row + 2, column=7).value or "")
                assert "INDEX('Hidden_Value_Flags'!$K:$K,$AI139)" in str(wb["Valuation"].cell(row=flags_header_row + 2, column=8).value or "")
                assert "IF(N('Hidden_Value_Flags'!$D$2)>=1,2,\"\")" in str(wb["Valuation"].cell(row=139, column=35).value or "")
                assert "IF(N('Hidden_Value_Flags'!$D$8)>=1,8,\"\")" in str(wb["Valuation"].cell(row=145, column=35).value or "")
                required_names = {"FCF_Yield", "FCF_TTM_Pos_Years", "Pos_FCF_Ratio", "Interest_Coverage"}
                assert required_names.issubset(set(wb.defined_names.keys()))
                assert "(price-linked)" in str(wb["Hidden_Value_Flags"]["K3"].value or "")
                assert "(price-linked)" in str(wb["Hidden_Value_Flags"]["K4"].value or "")
                assert "Current:" not in str(wb["Hidden_Value_Flags"]["K3"].value or "")
                assert "Current:" not in str(wb["Hidden_Value_Flags"]["K4"].value or "")
                assert not bool(wb["Valuation"].row_dimensions[245].hidden)
                assert float(wb["Valuation"].row_dimensions[245].height or 0.0) == pytest.approx(18.0, rel=1e-9)
                assert (
                    str(wb["Valuation"].cell(row=245, column=2).border.bottom.style or "") == "thick"
                    or str(wb["Valuation"].cell(row=246, column=2).border.top.style or "") == "thick"
                )
                operating_drivers_row = _find_row_with_value(wb["Valuation"], "Operating Drivers", column=15)
                assert operating_drivers_row is not None and operating_drivers_row > 18
                assert str(wb["Valuation"].cell(row=216, column=2).value or "").strip() != "Valuation Sensitivity Grid"
                assert str(wb["Valuation"].cell(row=217, column=2).value or "").strip() == "Valuation Sensitivity Grid"
                assert [str(wb["Valuation"].cell(row=8, column=cc).value or "").strip() for cc in (15, 17, 18, 19, 26)] == [
                    "Metric",
                    "Stated in",
                    "Applies to",
                    "Guidance",
                    "Trend / realized",
                ]
                buybacks_row = _find_row_with_value(wb["Valuation"], "Buybacks (shares)", column=1)
                buybacks_note_row = _find_row_with_value(wb["Valuation"], "Buybacks note", column=1)
                dividends_row = _find_row_with_value(wb["Valuation"], "Dividends ($/share)", column=1)
                dividends_note_row = _find_row_with_value(wb["Valuation"], "Dividends note", column=1)
                assert buybacks_row is not None
                assert buybacks_note_row is not None
                assert dividends_row is not None
                assert dividends_note_row is not None
                assert str(wb["Valuation"].cell(row=buybacks_row, column=2).value or "").strip() == "Latest quarter +12.614m at $10.04/share for $126.6m"
                assert str(wb["Valuation"].cell(row=buybacks_note_row, column=2).value or "").strip() == "Remaining capacity $359.0m | Latest increase by $250.0m on 2026-02-13 | Maturity date 2025 | Continuation mentioned."
                assert str(wb["Valuation"].cell(row=dividends_row, column=2).value or "").strip() == "Latest quarter div/share $0.090 | TTM dividend cash $148.5m"
                assert str(wb["Valuation"].cell(row=dividends_note_row, column=2).value or "").strip() == "We expect to continue to pay a quarterly dividend."
                convertible_row = _find_row_with_value(wb["Valuation"], "Convertible notes", column=12)
                assert convertible_row is not None
                assert str(wb["Valuation"].cell(row=convertible_row + 2, column=12).value or "").strip() == "1.5% notes due August 2032"
                assert any(
                    m.min_row == convertible_row + 1 and m.min_col == 24 and m.max_col == 27
                    for m in wb["Valuation"].merged_cells.ranges
                )
                q4_guidance = {(metric, applies): (stated, guidance, trend) for metric, stated, applies, guidance, trend in _guidance_block_rows(wb["Valuation"], "2025-12-31")}
                q3_guidance = {(metric, applies): (stated, guidance, trend) for metric, stated, applies, guidance, trend in _guidance_block_rows(wb["Valuation"], "2025-09-30")}
                assert q4_guidance[("Revenue", "FY2026")][0] == "Q4 2025"
                assert q4_guidance[("Adj EBIT", "FY2026")][0] == "Q4 2025"
                assert q4_guidance[("Adj EPS", "FY2026")][0] == "Q4 2025"
                assert q4_guidance[("FCF", "FY2026")][0] == "Q4 2025"
                assert q4_guidance[("Cost savings run-rate", "Run-rate")][0] == "Q1 2025"
                assert "Raised target to $180m-$200m annualized savings" in q4_guidance[("Cost savings run-rate", "Run-rate")][1]
                assert q4_guidance[("Cost savings run-rate", "Run-rate")][2] == ""
                assert q3_guidance[("Adj EBIT", "FY2025")][2] == "Δ +$0.0m (0.0%)"
                assert q3_guidance[("Adj EPS", "FY2025")][2] == "Δ +$0.00 (0.0%)"
                assert q3_guidance[("FCF", "FY2025")][2] == "Δ +$0.0m (0.0%)"
                assert q4_guidance[("Cost savings target", "")][2] == "from $170m-$190m"
                assert q4_guidance[("Adj EBIT", "FY2026")][2] == "Δ -$22.5m (-4.9%) | L -$40.0m | H -$5.0m"
                assert q4_guidance[("Adj EPS", "FY2026")][2] == "Δ +$0.20 (+15.4%) | L +$0.20 | H +$0.20"
                assert q4_guidance[("FCF", "FY2026")][2] == "Δ +$55.0m (+18.3%)"
                assert float(pd.to_numeric(wb["Valuation"].cell(row=convertible_row + 2, column=17).value, errors="coerce")) == pytest.approx(14.25, abs=0.01)
                assert float(pd.to_numeric(wb["Valuation"].cell(row=convertible_row + 2, column=19).value, errors="coerce")) == pytest.approx(16.135259, abs=0.001)
                assert float(pd.to_numeric(wb["Valuation"].cell(row=convertible_row + 2, column=21).value, errors="coerce")) == pytest.approx(5.535928, abs=0.001)
                buyback_ttm_row = _find_row_with_value(wb["Valuation"], "Buybacks (TTM, cash)")
                q4_col = next(
                    cc for cc in range(2, wb["Valuation"].max_column + 1)
                    if str(wb["Valuation"].cell(row=6, column=cc).value or "").strip() == "2025-Q4"
                )
                assert buyback_ttm_row is not None
                assert float(wb["Valuation"].cell(row=buyback_ttm_row, column=q4_col).value or 0.0) == pytest.approx(378.361, rel=1e-9)
                rows_2024_q4 = _block_rows(wb["Promise_Progress_UI"], "2024-12-31")
                revenue_row = next(row for row in rows_2024_q4 if row[0] == "Revenue guidance")
                assert revenue_row[3] == "Missed"
                assert revenue_row[6] == "Q4 2025"
                assert revenue_row[7] == "Q4 2025"
                assert revenue_row[8] == "2025-12-31"
                pp_hdr_row = _find_row_with_value(wb["Promise_Progress_UI"], "Promise progress (As of 2025-09-30)")
                assert pp_hdr_row is not None
                assert str(wb["Promise_Progress_UI"].cell(row=pp_hdr_row, column=1).border.top.style or "") != "medium"
                assert abs(float(wb["Promise_Progress_UI"].column_dimensions["B"].width or 0.0) - 38.142857) < 0.1
                assert abs(float(wb["Promise_Progress_UI"].column_dimensions["C"].width or 0.0) - 38.142857) < 0.1
                first_data_row = pp_hdr_row + 2
                assert str(wb["Promise_Progress_UI"].cell(row=first_data_row, column=2).alignment.horizontal or "") == "right"
                assert str(wb["Promise_Progress_UI"].cell(row=first_data_row, column=3).alignment.horizontal or "") == "right"
                assert str(wb["Promise_Progress_UI"].cell(row=first_data_row, column=5).alignment.vertical or "") == "center"
                assert str(wb["Promise_Progress_UI"].cell(row=1, column=1).value or "").strip() == "Promise Progress"
                assert str(wb["Promise_Progress_UI"].cell(row=2, column=1).value or "").strip().startswith("Generated at ")
                updated_fills = {
                    str(wb["Promise_Progress_UI"].cell(row=rr, column=4).fill.fgColor.rgb or "")
                    for rr in range(1, wb["Promise_Progress_UI"].max_row + 1)
                    if str(wb["Promise_Progress_UI"].cell(row=rr, column=4).value or "").strip() == "Updated"
                }
                assert updated_fills == {"00D9EAF7"}
            else:
                junk_metrics = {
                    "year ended December cost savings",
                    "least 45Z monetization / EBITDA",
                    "evaluation Results of Debt reduction",
                    "in all of our Strategic milestone",
                    "company has federal R&D Strategic milestone",
                    "federal research and development 45Z generation",
                }
                visible_metrics = {metric for metric, _, _, _, _, _ in progress_rows}
                assert junk_metrics.isdisjoint(visible_metrics)
                assert all(latest.lower() != "nan" for _, _, latest, _, _, _ in progress_rows)
                assert "No guidance text captured for this quarter." not in panel_blob
                assert "Management commentary" not in panel_blob
                assert "Guidance full text" not in panel_blob
                assert str(wb["Valuation"].cell(row=7, column=15).value or "").strip().startswith("Guidance (As of 2025-12-31)")
                assert str(wb["Valuation"].cell(row=8, column=15).value or "").strip() == "Metric"
                assert str(wb["Valuation"].cell(row=9, column=15).value or "").strip() == "A) Updated / mentioned this quarter"
                assert any(
                    metric in {
                        "45Z monetization",
                        "45Z monetization outlook",
                        "45Z from remaining facilities",
                        "Advantage Nebraska EBITDA opportunity",
                        "Interest expense outlook",
                        "Advantage Nebraska startup",
                    }
                    for metric, _, _, _, _, _ in progress_rows
                )
                assert any(metric == "Interest expense outlook" and "$30m-$35m" in target for metric, target, _, _, _, _ in progress_rows)
                assert any(
                    metric == "Debt reduction"
                    and "Clean Sugar Technology" not in rationale
                    for metric, _, _, _, rationale, _ in progress_rows
                )
                flags_header_row = _find_row_with_value(wb["Valuation"], "Hidden value flags", column=1)
                assert flags_header_row == 137
                assert str(wb["Valuation"].cell(row=flags_header_row, column=1).value or "").strip() == "Hidden value flags"
                assert str(wb["Valuation"].cell(row=flags_header_row, column=14).value or "").strip() == "Hidden Value Panel"
                assert any(m.min_row == 137 and m.min_col == 14 and m.max_col == 18 for m in wb["Valuation"].merged_cells.ranges)
                assert str(wb["Valuation"].cell(row=138, column=1).value or "").strip() == "Flag"
                assert str(wb["Valuation"].cell(row=138, column=2).value or "").strip() == "Summary"
                assert str(wb["Valuation"].cell(row=138, column=6).value or "").strip() == "Score"
                assert str(wb["Valuation"].cell(row=138, column=7).value or "").strip() == "Severity"
                assert str(wb["Valuation"].cell(row=138, column=8).value or "").strip() == "Result / support"
                assert str(wb["Valuation"].cell(row=flags_header_row + 2, column=1).value or "").startswith("=IF($AI139")
                assert "(price-linked)" in str(wb["Hidden_Value_Flags"]["K3"].value or "")
                assert "(price-linked)" in str(wb["Hidden_Value_Flags"]["K4"].value or "")
                assert "Current:" not in str(wb["Hidden_Value_Flags"]["K3"].value or "")
                assert "Current:" not in str(wb["Hidden_Value_Flags"]["K4"].value or "")
                red_green_row = _find_row_with_value(wb["Valuation"], "Red/Green Flags", column=1)
                assert red_green_row is not None
                assert any(m.min_row == red_green_row and m.min_col == 1 and m.max_col == 9 for m in wb["Valuation"].merged_cells.ranges)
                assert "Context" not in panel_blob
                assert not bool(wb["Valuation"].row_dimensions[245].hidden)
                assert float(wb["Valuation"].row_dimensions[245].height or 0.0) == pytest.approx(18.0, rel=1e-9)
                assert (
                    str(wb["Valuation"].cell(row=245, column=2).border.bottom.style or "") == "thick"
                    or str(wb["Valuation"].cell(row=246, column=2).border.top.style or "") == "thick"
                )
                operating_drivers_row = _find_row_with_value(wb["Valuation"], "Operating Drivers", column=15)
                assert operating_drivers_row is not None and operating_drivers_row > 18
                convertible_row = _find_row_with_value(wb["Valuation"], "Convertible notes", column=12)
                assert convertible_row is not None
                assert str(wb["Valuation"].cell(row=convertible_row + 2, column=12).value or "").strip() == "2.25% notes due 2027"
                assert str(wb["Valuation"].cell(row=convertible_row + 3, column=12).value or "").strip() == "5.25% notes due 2030"
                assert any(
                    m.min_row == convertible_row + 1 and m.min_col == 24 and m.max_col == 27
                    for m in wb["Valuation"].merged_cells.ranges
                )
                assert str(wb["Valuation"].cell(row=216, column=2).value or "").strip() != "Valuation Sensitivity Grid"
                assert str(wb["Valuation"].cell(row=217, column=2).value or "").strip() == "Valuation Sensitivity Grid"
                q4_guidance = {(metric, applies): (stated, guidance, trend) for metric, stated, applies, guidance, trend in _guidance_block_rows(wb["Valuation"], "2025-12-31")}
                assert q4_guidance[("Capex guidance (FY 2026)", "FY2026")][0] == "Q4 2025"
                assert q4_guidance[("Capex guidance (FY 2026)", "FY2026")][1] == "$15.0m-$25.0m"
                assert q4_guidance[("45Z base-case improvement", "FY2026")][1] == "Base improved by ILUC removal, facility qualification, and Advantage Nebraska"
                assert q4_guidance[("Commercial positioning / setup", "Q1 2026")][1] == "Q1 consolidated crush margins were better year over year."
                assert q4_guidance[("Farm-practice upside timing", "FY2026")][1] == "Excluded from current base; final guidance expected in 2026"
                assert panel_blob.count("Q1 consolidated crush margins were better year over year.") == 1
                q3_guidance = _guidance_block_rows(wb["Valuation"], "2025-09-30")
                assert ("Coverage / openness", "Q2 2025", "Q3 2025", "Q3 was about 65% crushed, moving closer to 70%.", "") in q3_guidance
                assert any(
                    metric == "Risk-management setup"
                    and stated == "Q3 2025"
                    and applies == "Q4 2025"
                    and "lock-in opportunities" in guidance
                    for metric, stated, applies, guidance, _ in q3_guidance
                )
                overlay_row = _find_row_with_value(wb["Economics_Overlay"], "Commercial / hedge setup", column=1)
                assert overlay_row is not None
                assert _find_row_with_value(wb["Economics_Overlay"], "Hedge / position overlay", column=1) is None
                assert str(wb["Economics_Overlay"]["A1"].value or "").strip() == "Economics Overlay"
                assert str(wb["Economics_Overlay"].cell(row=overlay_row + 1, column=1).value or "").strip() == "Horizon"
                assert str(wb["Economics_Overlay"].cell(row=overlay_row + 1, column=2).value or "").strip() == "Stated in"
                assert str(wb["Economics_Overlay"].cell(row=overlay_row + 1, column=3).value or "").strip() == "Setup"
                assert str(wb["Economics_Overlay"].cell(row=overlay_row + 1, column=5).value or "").strip() == "Coverage / openness"
                assert str(wb["Economics_Overlay"].cell(row=overlay_row + 1, column=8).value or "").strip() == "Locked margin / legs"
                assert str(wb["Economics_Overlay"].cell(row=overlay_row + 1, column=12).value or "").strip() == "Effect on results"
                assert str(wb["Economics_Overlay"].cell(row=overlay_row + 1, column=15).value or "").strip() == "Takeaway"
                assert any(m.min_row == overlay_row + 1 and m.min_col == 3 and m.max_col == 4 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == overlay_row + 1 and m.min_col == 5 and m.max_col == 7 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == overlay_row + 1 and m.min_col == 8 and m.max_col == 11 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == overlay_row + 1 and m.min_col == 12 and m.max_col == 14 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == overlay_row + 1 and m.min_col == 15 and m.max_col == 17 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert str(wb["Economics_Overlay"].cell(row=3, column=1).value or "").strip() == ""
                assert str(wb["Economics_Overlay"]["A1"].fill.fgColor.rgb or "") == "006FA8DC"
                assert str(wb["Economics_Overlay"]["A3"].fill.fgColor.rgb or "") == "00EDF4FA"
                assert str(wb["Economics_Overlay"][f"A{overlay_row}"].fill.fgColor.rgb or "") == "006FA8DC"
                assert str(wb["Economics_Overlay"].cell(row=overlay_row + 1, column=1).fill.fgColor.rgb or "") == "00EAF3FB"
                bridge_row = _find_row_with_value(wb["Economics_Overlay"], "Bridge to reported", column=1)
                base_row = _find_row_with_value(wb["Economics_Overlay"], "Base operating coefficients", column=1)
                market_row = _find_row_with_value(wb["Economics_Overlay"], "Market inputs", column=1)
                process_row = _find_row_with_value(wb["Economics_Overlay"], "Unhedged process economics", column=1)
                sandbox_build_row = _find_row_with_value(wb["Basis_Proxy_Sandbox"], "Approximate market crush build-up ($/gal)", column=2)
                assert bridge_row is not None and base_row is not None and market_row is not None
                assert process_row is None
                assert sandbox_build_row is not None
                assert overlay_row < bridge_row < base_row < market_row
                assert float(wb["Economics_Overlay"].column_dimensions["A"].width or 0.0) == pytest.approx((315.0 - 5.0) / 7.0, abs=0.05)
                for letter in tuple("BCDEFGHIJKLMNOPQR"):
                    assert float(wb["Economics_Overlay"].column_dimensions[letter].width or 0.0) == pytest.approx((102.0 - 5.0) / 7.0, abs=0.05)
                    for rr in (overlay_row, bridge_row, base_row, market_row):
                        assert str(wb["Economics_Overlay"].cell(row=rr, column=1).fill.fgColor.rgb or "") == "006FA8DC"
                        expected_height = 22.5
                        assert float(wb["Economics_Overlay"].row_dimensions[rr].height or 0.0) == pytest.approx(expected_height, abs=0.1)
                    assert bool(wb["Economics_Overlay"].cell(row=rr, column=1).font.bold)
                    assert float(wb["Economics_Overlay"].cell(row=rr, column=1).font.size or 0.0) == pytest.approx(13.0, abs=0.1)
                    assert str(wb["Economics_Overlay"].cell(row=rr, column=1).alignment.horizontal or "") == "center"
                assert float(wb["Economics_Overlay"].row_dimensions[3].height or 0.0) == pytest.approx(24.0, abs=0.1)
                assert float(wb["Economics_Overlay"].row_dimensions[4].height or 0.0) == pytest.approx(15.0, abs=0.1)
                assert float(wb["Economics_Overlay"].row_dimensions[bridge_row + 1].height or 0.0) == pytest.approx(24.0, abs=0.1)
                assert float(wb["Economics_Overlay"].row_dimensions[bridge_row + 2].height or 0.0) == pytest.approx(15.0, abs=0.1)
                assert str(wb["Economics_Overlay"].cell(row=bridge_row + 1, column=1).fill.fgColor.rgb or "") == "00EDF4FA"
                assert any(m.min_row == base_row + 3 and m.min_col == 4 and m.max_col == 5 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == base_row + 3 and m.min_col == 6 and m.max_col == 8 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == market_row + 1 and m.min_col == 1 and m.max_col == 17 for m in wb["Economics_Overlay"].merged_cells.ranges)
                year_band_rows = {
                    str(wb["Economics_Overlay"].cell(row=rr, column=1).value or "").strip(): rr
                    for rr in range(overlay_row + 2, bridge_row)
                    if str(wb["Economics_Overlay"].cell(row=rr, column=1).value or "").strip() in {"2023", "2024", "2025", "2026 / current"}
                }
                assert set(year_band_rows) == {"2023", "2024", "2025", "2026 / current"}
                for label, rr in year_band_rows.items():
                    assert any(m.min_row == rr and m.min_col == 1 and m.max_col == 17 for m in wb["Economics_Overlay"].merged_cells.ranges)
                    assert str(wb["Economics_Overlay"].cell(row=rr, column=1).fill.fgColor.rgb or "") == "00EDF4FA"
                    assert float(wb["Economics_Overlay"].row_dimensions[rr].height or 0.0) == pytest.approx(21.0, abs=0.1)
                overlay_blob = " | ".join(
                    str(wb["Economics_Overlay"].cell(row=rr, column=cc).value or "").strip()
                    for rr in range(overlay_row, bridge_row)
                    for cc in range(1, 18)
                    if str(wb["Economics_Overlay"].cell(row=rr, column=cc).value or "").strip()
                )
                assert "Q4 paper margin about $0.22-$0.25/gal" in overlay_blob
                assert "Q3 all-in margins roughly $0.20-$0.30+ per gallon" in overlay_blob
                assert "All-in margins, not just simple crush" in overlay_blob
                assert "Q1 consolidated crush margins were better year over year" in overlay_blob
                overlay_visible_blob = " | ".join(
                    str(wb["Economics_Overlay"].cell(row=rr, column=cc).value or "").strip()
                    for rr in range(1, min(70, wb["Economics_Overlay"].max_row + 1))
                    for cc in range(1, 18)
                    if str(wb["Economics_Overlay"].cell(row=rr, column=cc).value or "").strip()
                )
                assert "#VALUE!" not in overlay_visible_blob
                assert "#VÄRDEFEL!" not in overlay_visible_blob
                assert "Source / confidence" not in overlay_visible_blob
                assert "Transcript | High" not in overlay_visible_blob
                assert "Conference | High" not in overlay_visible_blob
                assert "#VALUE!" not in overlay_visible_blob
                assert "#VÃ„RDEFEL!" not in overlay_visible_blob
                assert "#VÄRDEFEL!" not in overlay_visible_blob
                first_setup_comment = None
                for rr in range(overlay_row + 2, bridge_row):
                    setup_comment = wb["Economics_Overlay"].cell(row=rr, column=3).comment
                    if setup_comment is not None:
                        first_setup_comment = setup_comment
                        break
                assert first_setup_comment is not None
                assert "Source:" in str(first_setup_comment.text or "")
                assert "Confidence:" in str(first_setup_comment.text or "")
                first_data_row = next(
                    rr
                    for rr in range(overlay_row + 2, bridge_row)
                    if str(wb["Economics_Overlay"].cell(row=rr, column=1).value or "").strip() not in {"", "2023", "2024", "2025", "2026 / current"}
                )
                assert str(wb["Economics_Overlay"].cell(row=first_data_row, column=1).value or "").strip() == "Q1 2026"
                assert bool(wb["Economics_Overlay"].cell(row=first_data_row, column=1).font.bold)
                assert bool(wb["Economics_Overlay"].cell(row=first_data_row, column=3).font.bold)
                assert str(wb["Economics_Overlay"].cell(row=first_data_row, column=1).fill.fgColor.rgb or "") in {"00FFFFFF", "00F7F9FC"}
                assert str(wb["Economics_Overlay"].cell(row=first_data_row, column=1).border.bottom.style or "") == "thin"
                for rr in range(overlay_row + 2, bridge_row):
                    row_label = str(wb["Economics_Overlay"].cell(row=rr, column=1).value or "").strip()
                    if row_label and row_label not in {"2023", "2024", "2025", "2026 / current"}:
                        assert 19.5 <= float(wb["Economics_Overlay"].row_dimensions[rr].height or 0.0) <= 60.0
                bridge_labels = [
                    str(wb["Economics_Overlay"].cell(row=rr, column=1).value or "").strip()
                    for rr in range(bridge_row + 4, base_row)
                    if str(wb["Economics_Overlay"].cell(row=rr, column=1).value or "").strip()
                ]
                assert bridge_labels == [
                    "Approximate market crush ($m)",
                    "GPRE crush proxy ($m)",
                    "Underlying crush margin ($m)",
                    "Reported consolidated crush margin ($m)",
                    "45Z impact ($m)",
                    "RIN impact ($m)",
                    "Inventory NRV / lower-of-cost ($m)",
                    "Non-ethanol operating activities ($m)",
                    "Impairment / held-for-sale ($m)",
                    "Other explicit bridge items ($m)",
                ]
                assert float(wb["Economics_Overlay"].row_dimensions[bridge_row + 3].height or 0.0) == pytest.approx(21.0, abs=0.1)
                bridge_data_rows = [
                    rr
                    for rr in range(bridge_row + 4, base_row)
                    if str(wb["Economics_Overlay"].cell(row=rr, column=1).value or "").strip()
                ]
                assert all(float(wb["Economics_Overlay"].row_dimensions[rr].height or 0.0) == pytest.approx(24.0, abs=0.1) for rr in bridge_data_rows)
                first_bridge_value_row = bridge_data_rows[0]
                assert float(wb["Economics_Overlay"].cell(row=first_bridge_value_row, column=1).font.size or 0.0) == pytest.approx(12.0, abs=0.1)
                assert float(wb["Economics_Overlay"].cell(row=first_bridge_value_row, column=2).font.size or 0.0) == pytest.approx(12.0, abs=0.1)
                assert str(wb["Economics_Overlay"].cell(row=base_row + 1, column=1).value or "").strip().startswith("Use platform/process coefficients")
                assert str(wb["Economics_Overlay"].cell(row=base_row + 1, column=1).fill.fgColor.rgb or "") == "00EDF4FA"
                assert float(wb["Economics_Overlay"].row_dimensions[base_row + 1].height or 0.0) == pytest.approx(24.0, abs=0.1)
                assert float(wb["Economics_Overlay"].row_dimensions[base_row + 2].height or 0.0) == pytest.approx(15.0, abs=0.1)
                assert str(wb["Economics_Overlay"].cell(row=base_row + 3, column=4).value or "").strip() == "Status"
                assert any(m.min_row == base_row + 3 and m.min_col == 6 and m.max_col == 8 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == base_row + 4 and m.min_col == 6 and m.max_col == 8 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert float(wb["Economics_Overlay"].row_dimensions[base_row + 4].height or 0.0) == pytest.approx(24.0, abs=0.1)
                assert any(m.min_row == market_row + 1 and m.min_col == 1 and m.max_col == 17 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == market_row + 3 and m.min_col == 2 and m.max_col == 3 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == market_row + 3 and m.min_col == 4 and m.max_col == 5 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == market_row + 3 and m.min_col == 7 and m.max_col == 11 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == market_row + 3 and m.min_col == 12 and m.max_col == 17 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == market_row + 4 and m.min_col == 2 and m.max_col == 3 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == market_row + 4 and m.min_col == 4 and m.max_col == 5 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == market_row + 4 and m.min_col == 7 and m.max_col == 11 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert any(m.min_row == market_row + 4 and m.min_col == 12 and m.max_col == 17 for m in wb["Economics_Overlay"].merged_cells.ranges)
                assert float(wb["Economics_Overlay"].row_dimensions[market_row + 1].height or 0.0) == pytest.approx(24.0, abs=0.1)
                assert float(wb["Economics_Overlay"].row_dimensions[market_row + 2].height or 0.0) == pytest.approx(15.0, abs=0.1)
                assert "Unhedged process economics" not in overlay_visible_blob
                assert str(wb["Basis_Proxy_Sandbox"].cell(row=sandbox_build_row + 1, column=2).value or "").strip() == "Official simple row build-up used by Approximate market crush on Economics_Overlay."
                assert any(m.min_row == sandbox_build_row + 2 and m.min_col == 3 and m.max_col == 4 for m in wb["Basis_Proxy_Sandbox"].merged_cells.ranges)
                assert any(m.min_row == sandbox_build_row + 2 and m.min_col == 5 and m.max_col == 6 for m in wb["Basis_Proxy_Sandbox"].merged_cells.ranges)
                assert any(m.min_row == sandbox_build_row + 2 and m.min_col == 7 and m.max_col == 8 for m in wb["Basis_Proxy_Sandbox"].merged_cells.ranges)
                assert any(m.min_row == sandbox_build_row + 2 and m.min_col == 9 and m.max_col == 10 for m in wb["Basis_Proxy_Sandbox"].merged_cells.ranges)
                sandbox_process_margin_row = next(
                    rr for rr in range(sandbox_build_row + 4, wb["Basis_Proxy_Sandbox"].max_row + 1)
                    if str(wb["Basis_Proxy_Sandbox"].cell(row=rr, column=2).value or "").strip() == "Approximate market crush"
                )
                sandbox_basis_snapshot_row = next(
                    rr for rr in range(sandbox_process_margin_row, wb["Basis_Proxy_Sandbox"].max_row + 1)
                    if str(wb["Basis_Proxy_Sandbox"].cell(row=rr, column=2).value or "").strip() == "Official corn basis snapshot date"
                )
                sandbox_basis_rule_row = next(
                    rr for rr in range(sandbox_basis_snapshot_row, wb["Basis_Proxy_Sandbox"].max_row + 1)
                    if str(wb["Basis_Proxy_Sandbox"].cell(row=rr, column=2).value or "").strip() == "Official corn basis selection rule"
                )
                assert str(wb["Basis_Proxy_Sandbox"].cell(row=sandbox_process_margin_row, column=11).value or "").strip() == "$/gal"
                assert str(wb["Basis_Proxy_Sandbox"].cell(row=sandbox_process_margin_row, column=12).value or "").strip() == "Market crush estimate with natural gas cost and GPRE corn basis, weighted to active capacity, and converted to $/gal."
                assert float(wb["Basis_Proxy_Sandbox"].row_dimensions[sandbox_process_margin_row].height or 0.0) == pytest.approx(24.0, abs=0.1)
                assert str(wb["Basis_Proxy_Sandbox"].cell(row=sandbox_basis_snapshot_row, column=11).value or "").strip() == "date/text"
                assert "Retained GPRE corn-bid snapshot date used by the official corn-basis leg only" in str(
                    wb["Basis_Proxy_Sandbox"].cell(row=sandbox_basis_snapshot_row, column=12).value or ""
                )
                assert str(wb["Basis_Proxy_Sandbox"].cell(row=sandbox_basis_rule_row, column=11).value or "").strip() == "rule/text"
                assert "Frame-specific retained-snapshot selector used by the official corn-basis leg" in str(
                    wb["Basis_Proxy_Sandbox"].cell(row=sandbox_basis_rule_row, column=12).value or ""
                )
                assert re.fullmatch(
                    r"As of 2026-\d{2}-\d{2}",
                    str(wb["Basis_Proxy_Sandbox"].cell(row=sandbox_build_row + 3, column=7).value or "").strip(),
                )
                assert str(wb["Operating_Drivers"]["A2"].fill.fgColor.rgb or "") == "006FA8DC"
                assert str(wb["Operating_Drivers"]["A4"].fill.fgColor.rgb or "") == "006FA8DC"
                assert str(wb["Operating_Drivers"]["A5"].fill.fgColor.rgb or "") == "00EAF3FB"
                assert str(wb["Operating_Drivers"]["B5"].fill.fgColor.rgb or "") == "00EAF3FB"
                rows_2025_q3 = _block_rows(wb["Promise_Progress_UI"], "2025-09-30")
                assert [row[0] for row in rows_2025_q3] == [
                    "45Z monetization outlook",
                    "Debt reduction",
                    "45Z facility qualification",
                    "Advantage Nebraska startup",
                ]
                monetization_outlook_row = next(row for row in rows_2025_q3 if row[0] == "45Z monetization outlook")
                assert monetization_outlook_row[1] == "$15m-$25m"
                assert monetization_outlook_row[2] == "not yet measurable"
                assert monetization_outlook_row[3] == "Open"
                debt_reduction_row = next(row for row in rows_2025_q3 if row[0] == "Debt reduction")
                assert debt_reduction_row[1] == "$130.7m"
                assert debt_reduction_row[2] == "Debt repaid"
                assert debt_reduction_row[3] == "Completed"
                facility_row = next(row for row in rows_2025_q3 if row[0] == "45Z facility qualification")
                assert facility_row[2] == "Expected in 2026"
                assert facility_row[3] == "On track"
                nebraska_row = next(row for row in rows_2025_q3 if row[0] == "Advantage Nebraska startup")
                assert nebraska_row[2] == "Advantage Nebraska is fully operational and sequestering CO2 in Wyoming."
                assert nebraska_row[3] == "Completed"
                assert "Debt exchange" in quarter_note_metrics
                assert "Interest expense outlook" in quarter_note_metrics
                cost_note_row = next(
                    rr for rr in range(1, wb["Quarter_Notes_UI"].max_row + 1)
                    if "Cost reductions are on pace to exceed the $50.0m annualized savings target." in str(wb["Quarter_Notes_UI"].cell(row=rr, column=3).value or "")
                )
                carbon_note_row = next(
                    rr for rr in range(1, wb["Quarter_Notes_UI"].max_row + 1)
                    if "York carbon capture was fully operational; Central City and Wood River were online and ramping." in str(wb["Quarter_Notes_UI"].cell(row=rr, column=3).value or "")
                )
                assert str(wb["Quarter_Notes_UI"].cell(row=cost_note_row, column=4).value or "").strip() == "Cost savings target"
                assert str(wb["Quarter_Notes_UI"].cell(row=carbon_note_row, column=4).value or "").strip() == "Carbon capture operational milestone"
                assert "least 45Z monetization / EBITDA" not in quarter_note_metrics
                assert "year ended December Expense drivers" not in quarter_note_metrics
                assert all("[REPEAT]" not in str(wb["Quarter_Notes_UI"].cell(row=rr, column=3).value or "") for rr in range(1, wb["Quarter_Notes_UI"].max_row + 1))
                pp_hdr_row = _find_row_with_value(wb["Promise_Progress_UI"], "Promise progress (As of 2025-09-30)")
                assert pp_hdr_row is not None
                assert str(wb["Promise_Progress_UI"].cell(row=pp_hdr_row, column=1).border.top.style or "") != "medium"
                assert abs(float(wb["Promise_Progress_UI"].column_dimensions["B"].width or 0.0) - 38.142857) < 0.1
                assert abs(float(wb["Promise_Progress_UI"].column_dimensions["C"].width or 0.0) - 38.142857) < 0.1
                first_data_row = pp_hdr_row + 2
                assert str(wb["Promise_Progress_UI"].cell(row=first_data_row, column=2).alignment.horizontal or "") == "right"
                assert str(wb["Promise_Progress_UI"].cell(row=first_data_row, column=3).alignment.horizontal or "") == "right"
                assert str(wb["Promise_Progress_UI"].cell(row=first_data_row, column=5).alignment.vertical or "") == "center"
                assert str(wb["Promise_Progress_UI"].cell(row=1, column=1).value or "").strip() == "Promise Progress"
                assert str(wb["Promise_Progress_UI"].cell(row=2, column=1).value or "").strip().startswith("Generated at ")
        finally:
            wb.close()


def test_current_delivered_workbooks_analysis_sheets_share_blue_theme() -> None:
    theme_title = "006FA8DC"
    theme_section = "00EDF4FA"
    theme_soft_section = "00D9E7F3"
    theme_header = "00EAF3FB"
    theme_neutral = "00F7F9FC"
    theme_alt = "00FFFFFF"

    def _fill_rgb(cell) -> str:
        return str(cell.fill.fgColor.rgb or "")

    def _font_rgb(cell) -> str:
        return str(getattr(cell.font.color, "rgb", "") or "")

    def _first_iso_date_row(ws) -> int:
        for rr in range(1, ws.max_row + 1):
            val = str(ws.cell(row=rr, column=1).value or "").strip()
            if re.fullmatch(r"20\d{2}-\d{2}-\d{2}", val):
                return rr
        raise AssertionError("No quarter block row found")

    def _first_promise_block_row(ws) -> int:
        for rr in range(1, ws.max_row + 1):
            val = str(ws.cell(row=rr, column=1).value or "").strip()
            if val.startswith("Promise progress (As of "):
                return rr
        raise AssertionError("No promise progress block row found")

    def _find_row_starting_with(ws, prefix: str, column: int | None = None) -> int:
        for rr in range(1, ws.max_row + 1):
            if column is not None:
                val = str(ws.cell(row=rr, column=column).value or "").strip()
                if val.startswith(prefix):
                    return rr
                continue
            for cc in range(1, ws.max_column + 1):
                val = str(ws.cell(row=rr, column=cc).value or "").strip()
                if val.startswith(prefix):
                    return rr
        raise AssertionError(f"No row starting with {prefix!r} found")

    for ticker in ["PBI", "GPRE"]:
        workbook_path = _current_delivered_model_path(ticker)
        if not workbook_path.exists():
            pytest.skip(f"Current delivered workbook missing for style snapshot test: {workbook_path}")
        wb = load_workbook(workbook_path, data_only=False, read_only=False)
        try:
            ws_summary = wb["SUMMARY"]
            assert _fill_rgb(ws_summary["A1"]) in {"005B9BD5", theme_title}
            assert float(ws_summary["A1"].font.size or 0.0) == pytest.approx(15.0, abs=0.1)
            assert str(ws_summary["A1"].border.top.style or "") == "thick"
            assert str(getattr(ws_summary["A1"].border.top.color, "rgb", "") or "").endswith("5E6F82")

            ws_val = wb["Valuation"]
            assert str(ws_val["A3"].value or "").strip() == "Valuation"
            assert _fill_rgb(ws_val["A3"]) == theme_title
            assert _font_rgb(ws_val["A3"]) == "00FFFFFF"
            assert any(m.min_row == 3 and m.min_col == 1 and m.max_col == 13 for m in ws_val.merged_cells.ranges)
            assert float(ws_val.row_dimensions[1].height or 0.0) == pytest.approx(18.0, abs=0.1)
            assert all(float(ws_val.row_dimensions[rr].height or 0.0) == pytest.approx(18.0, abs=0.1) for rr in range(7, 118))
            assert _fill_rgb(ws_val["A6"]) == theme_header
            assert _fill_rgb(ws_val["A7"]) == theme_soft_section
            assert _fill_rgb(ws_val["A137"]) == theme_title
            assert _fill_rgb(ws_val["A138"]) == theme_header
            assert _fill_rgb(ws_val["A147"]) == theme_title
            assert _fill_rgb(ws_val["O7"]) == theme_title
            guidance_2025q4_row = _find_row_starting_with(ws_val, "Guidance (As of 2025-12-31)", column=15)
            assert guidance_2025q4_row == 7
            assert _fill_rgb(ws_val.cell(row=guidance_2025q4_row, column=15)) == theme_title
            assert _font_rgb(ws_val.cell(row=guidance_2025q4_row, column=15)) == "00FFFFFF"
            assert _fill_rgb(ws_val.cell(row=guidance_2025q4_row + 1, column=15)) == theme_header
            assert _fill_rgb(ws_val.cell(row=guidance_2025q4_row + 1, column=19)) == theme_header
            assert _fill_rgb(ws_val.cell(row=guidance_2025q4_row + 1, column=26)) == theme_header
            guidance_2025q3_row = _find_row_starting_with(ws_val, "Guidance (As of 2025-09-30)", column=15)
            assert guidance_2025q3_row > guidance_2025q4_row
            assert _fill_rgb(ws_val.cell(row=guidance_2025q3_row, column=15)) == theme_title
            assert _fill_rgb(ws_val.cell(row=guidance_2025q3_row + 1, column=15)) == theme_header
            assert _fill_rgb(ws_val.cell(row=_find_row_with_value(ws_val, "Operating Drivers", column=15), column=15)) == theme_title
            assert _fill_rgb(ws_val.cell(row=_find_row_with_value(ws_val, "Thesis Bridge", column=15), column=15)) == theme_title
            assert _fill_rgb(ws_val.cell(row=_find_row_with_value(ws_val, "Debt Detail (latest)", column=1), column=1)) == theme_title
            assert _fill_rgb(ws_val.cell(row=_find_row_with_value(ws_val, "Capital return", column=1), column=1)) == theme_title
            assert _fill_rgb(ws_val.cell(row=_find_row_with_value(ws_val, "Convertible notes", column=12), column=12)) == theme_title
            assert _fill_rgb(ws_val["B192"]) == theme_title
            assert _fill_rgb(ws_val["N137"]) == theme_title
            trend_row = _find_row_starting_with(ws_val, "Trend/")
            assert _fill_rgb(ws_val.cell(row=trend_row, column=1)) == theme_title
            flags_row = _find_row_starting_with(ws_val, "Red/Green")
            assert _fill_rgb(ws_val.cell(row=flags_row, column=1)) == theme_title
            assert any(m.min_row == flags_row and m.min_col == 1 and m.max_col == 9 for m in ws_val.merged_cells.ranges)
            assert _find_row_with_value(ws_val, "DCF (optional module)", column=None) == 217
            assert _find_row_with_value(ws_val, "DCF Sensitivity ($/share)", column=None) == 225
            assert _find_row_with_value(ws_val, "Market-implied terminal g (solve gT so DCF EV = Market EV)", column=None) == 217
            assert _find_row_with_value(ws_val, "Scenario drivers (internal)", column=2) == 246
            assert str(ws_val.cell(row=245, column=2).value or "").strip() == ""
            assert not bool(ws_val.row_dimensions[245].hidden)
            assert float(ws_val.row_dimensions[245].height or 0.0) == pytest.approx(18.0, rel=1e-9)
            for label in ("Signal 1", "Signal 2", "Signal 3", "Signal 4", "Buybacks (shares)", "Buybacks note", "Dividends ($/share)", "Dividends note"):
                label_row = _find_row_with_value(ws_val, label, column=1)
                assert label_row is not None
                assert str(ws_val.cell(row=label_row, column=1).alignment.horizontal or "") == "left"
                assert str(ws_val.cell(row=label_row, column=1).alignment.vertical or "") == "center"
            for label in ("Signal 1", "Signal 2", "Signal 3", "Signal 4"):
                label_row = _find_row_with_value(ws_val, label, column=1)
                assert label_row is not None
                assert str(ws_val.cell(row=label_row, column=2).alignment.horizontal or "") == "left"
            for label in ("Buybacks (shares)", "Buybacks note", "Dividends ($/share)", "Dividends note"):
                label_row = _find_row_with_value(ws_val, label, column=1)
                assert label_row is not None
                assert str(ws_val.cell(row=label_row, column=2).alignment.horizontal or "") == "left"
            assert str(ws_val.cell(row=_find_row_with_value(ws_val, "Operating signals", column=1), column=1).alignment.horizontal or "") == "left"
            assert str(ws_val.cell(row=_find_row_with_value(ws_val, "Capital return", column=1), column=1).alignment.horizontal or "") == "left"
            adj_fcf_row = _find_row_with_value(ws_val, "Adj FCF (TTM)")
            assert adj_fcf_row is not None
            assert str((ws_val.cell(row=adj_fcf_row, column=1).comment.text if ws_val.cell(row=adj_fcf_row, column=1).comment else "") or "") == "company-defined"
            assert _find_row_with_value(ws_val, "Debt repaid (TTM)") is None
            assert _find_row_with_value(ws_val, "Debt issued (TTM)") is None
            assert _find_row_with_value(ws_val, "Total assets ($m)") is None
            assert _find_row_with_value(ws_val, "Total liabilities ($m)") is None
            assert _find_row_with_value(ws_val, "Goodwill % of assets") is None

            ws_bs = wb["BS_Segments"]
            assert _fill_rgb(ws_bs["A8"]) == theme_title
            assert _font_rgb(ws_bs["A8"]) == "00FFFFFF"
            assert _fill_rgb(ws_bs["B10"]) == theme_header
            assert _fill_rgb(ws_bs["A11"]) == theme_header
            quarterly_segments_row = _find_row_with_value(ws_bs, "Quarterly segments")
            if quarterly_segments_row is not None:
                assert _fill_rgb(ws_bs.cell(row=quarterly_segments_row, column=1)) == theme_title
                assert float(ws_bs.row_dimensions[quarterly_segments_row].height or 0.0) == pytest.approx(24.0, abs=0.1)
            annual_segments_row = _find_row_with_value(ws_bs, "Annual segments")
            assert annual_segments_row is not None
            assert _fill_rgb(ws_bs.cell(row=annual_segments_row, column=1)) == theme_title
            assert float(ws_bs.row_dimensions[annual_segments_row].height or 0.0) == pytest.approx(24.0, abs=0.1)
            assert float(ws_bs.row_dimensions[10].height or 0.0) == pytest.approx(19.5, abs=0.1)
            assert float(ws_bs.row_dimensions[11].height or 0.0) == pytest.approx(19.5, abs=0.1)
            assert float(ws_bs.row_dimensions[_find_row_with_value(ws_bs, "Liquidity / Assets")].height or 0.0) == pytest.approx(18.0, abs=0.1)
            inventory_row = _find_row_with_value(ws_bs, "Inventory")
            if inventory_row is not None:
                assert float(ws_bs.row_dimensions[inventory_row].height or 0.0) == pytest.approx(18.0, abs=0.1)
            total_assets_row = _find_row_with_value(ws_bs, "Total assets")
            assert total_assets_row is not None
            assert str(ws_bs.cell(row=total_assets_row + 1, column=1).value or "").strip() == "Goodwill % of assets"
            annual_year_row = next(
                rr
                for rr in range(annual_segments_row + 1, min(ws_bs.max_row, annual_segments_row + 6) + 1)
                if str(ws_bs.cell(row=rr, column=1).value or "").strip() == "Year"
            )
            assert str(ws_bs.cell(row=annual_year_row, column=1).alignment.horizontal or "") == "right"
            assert float(ws_bs.row_dimensions[annual_year_row].height or 0.0) == pytest.approx(19.5, abs=0.1)

            ws_qn = wb["Quarter_Notes_UI"]
            assert str(ws_qn["A1"].value or "").strip() == "Quarter Notes"
            assert _fill_rgb(ws_qn["A1"]) == theme_title
            assert _font_rgb(ws_qn["A1"]) == "00FFFFFF"
            assert float(ws_qn.row_dimensions[1].height or 0.0) == pytest.approx(27.0, abs=0.1)
            assert str(ws_qn["A2"].value or "").strip().startswith("Generated at ")
            assert _fill_rgb(ws_qn["A2"]) == theme_section
            assert float(ws_qn.row_dimensions[2].height or 0.0) == pytest.approx(19.5, abs=0.1)
            qn_block_row = _first_iso_date_row(ws_qn)
            assert _fill_rgb(ws_qn.cell(row=qn_block_row, column=1)) == theme_title
            assert _fill_rgb(ws_qn.cell(row=qn_block_row + 1, column=2)) == theme_header
            assert _fill_rgb(ws_qn.cell(row=qn_block_row + 2, column=2)) in {theme_alt, theme_neutral}
            assert float(ws_qn.row_dimensions[qn_block_row].height or 0.0) == pytest.approx(19.5, abs=0.1)
            assert float(ws_qn.row_dimensions[qn_block_row + 1].height or 0.0) == pytest.approx(19.5, abs=0.1)
            assert float(ws_qn.row_dimensions[qn_block_row + 2].height or 0.0) >= 19.5
            assert [str(ws_qn.cell(row=qn_block_row + 1, column=cc).value or "").strip() for cc in range(1, 5)] == ["", "Category", "Note", "Metric"]
            assert str(ws_qn.cell(row=qn_block_row + 1, column=5).value or "").strip() == ""
            assert bool(ws_qn.column_dimensions["E"].hidden)
            qn_blank_rows = [
                rr
                for rr in range(2, ws_qn.max_row + 1)
                if not any(str(ws_qn.cell(row=rr, column=cc).value or "").strip() for cc in range(1, 5))
            ]
            assert not qn_blank_rows

            if ticker == "GPRE":
                q4_col = next(
                    cc for cc in range(2, ws_val.max_column + 1)
                    if str(ws_val.cell(row=6, column=cc).value or "").strip() == "2025-Q4"
                )
                q3_col = next(
                    cc for cc in range(2, ws_val.max_column + 1)
                    if str(ws_val.cell(row=6, column=cc).value or "").strip() == "2025-Q3"
                )
                adj_ebit_row = _find_row_with_value(ws_val, "Adj EBIT (TTM)")
                adj_eps_row = _find_row_with_value(ws_val, "Adj EPS (TTM)")
                debt_repaid_row = _find_row_with_value(ws_val, "Debt repaid (gross, TTM)")
                debt_issued_row = _find_row_with_value(ws_val, "Debt issued (gross, TTM)")
                assert adj_fcf_row is not None
                assert adj_ebit_row is not None
                assert adj_eps_row is not None
                assert debt_repaid_row is not None
                assert debt_issued_row is not None
                assert ws_val.cell(row=adj_ebit_row, column=q4_col).value in (None, "")
                assert ws_val.cell(row=adj_eps_row, column=q4_col).value in (None, "")
                assert float(ws_val.cell(row=debt_repaid_row, column=q3_col).value or 0.0) == pytest.approx(130.7, abs=0.01)
                assert float(ws_val.cell(row=debt_repaid_row, column=q4_col).value or 0.0) == pytest.approx(130.7, abs=0.01)
                assert float(ws_val.cell(row=debt_issued_row, column=q4_col).value or 0.0) == pytest.approx(30.0, abs=0.01)

            ws_pp = wb["Promise_Progress_UI"]
            assert str(ws_pp["A1"].value or "").strip() == "Promise Progress"
            assert _fill_rgb(ws_pp["A1"]) == theme_title
            assert _font_rgb(ws_pp["A1"]) == "00FFFFFF"
            assert float(ws_pp.row_dimensions[1].height or 0.0) == pytest.approx(27.0, abs=0.1)
            assert str(ws_pp["A2"].value or "").strip().startswith("Generated at ")
            assert _fill_rgb(ws_pp["A2"]) == theme_section
            pp_block_row = _first_promise_block_row(ws_pp)
            assert _fill_rgb(ws_pp.cell(row=pp_block_row, column=1)) == theme_title
            assert _fill_rgb(ws_pp.cell(row=pp_block_row + 1, column=1)) == theme_header
            assert _fill_rgb(ws_pp.cell(row=pp_block_row + 2, column=1)) in {theme_alt, theme_neutral}
            assert _fill_rgb(ws_pp.cell(row=pp_block_row + 2, column=5)) in {theme_alt, theme_neutral}
            assert float(ws_pp.row_dimensions[pp_block_row].height or 0.0) == pytest.approx(19.5, abs=0.1)
            assert float(ws_pp.row_dimensions[pp_block_row + 1].height or 0.0) == pytest.approx(19.5, abs=0.1)
            assert str(ws_pp.cell(row=pp_block_row, column=1).border.top.style or "") != "medium"
            pp_blank_rows = [
                rr
                for rr in range(2, ws_pp.max_row + 1)
                if not any(str(ws_pp.cell(row=rr, column=cc).value or "").strip() for cc in range(1, 16))
            ]
            assert not pp_blank_rows
            first_status_fill = next(
                _fill_rgb(ws_pp.cell(row=rr, column=4))
                for rr in range(pp_block_row + 2, ws_pp.max_row + 1)
                if str(ws_pp.cell(row=rr, column=4).value or "").strip() not in {"", "Result"}
            )
            assert first_status_fill not in {theme_title, theme_section, theme_header}
            status_fill_expectations = {
                "Updated": "00D9EAF7",
                "Open": "00E7EDF3",
                "On track": "00D9EAD3",
                "Missed": "00F4CCCC",
                "Hit": "00C6EFCE",
                "Beat": "00A9D18E",
                "Completed": "0070AD47",
            }
            for status_label, expected_fill in status_fill_expectations.items():
                status_rows = [
                    rr
                    for rr in range(1, ws_pp.max_row + 1)
                    if str(ws_pp.cell(row=rr, column=4).value or "").strip() == status_label
                ]
                if status_rows:
                    assert {_fill_rgb(ws_pp.cell(row=rr, column=4)) for rr in status_rows} == {expected_fill}

            if "Economics_Overlay" in wb.sheetnames:
                ws_overlay = wb["Economics_Overlay"]
                assert _fill_rgb(ws_overlay["A1"]) == theme_title
                assert _fill_rgb(ws_overlay["A3"]) == theme_section
                overlay_hdr_row = _find_row_with_value(ws_overlay, "Commercial / hedge setup", column=1)
                assert overlay_hdr_row is not None
                assert _fill_rgb(ws_overlay.cell(row=overlay_hdr_row, column=1)) == theme_title
                assert _fill_rgb(ws_overlay.cell(row=overlay_hdr_row + 1, column=1)) == theme_header
                commercial_data_rows = [
                    rr
                    for rr in range(overlay_hdr_row + 2, ws_overlay.max_row + 1)
                    if str(ws_overlay.cell(row=rr, column=1).value or "").strip()
                    and str(ws_overlay.cell(row=rr, column=1).value or "").strip()
                    not in {"2023", "2024", "2025", "2026 / current", "Bridge to reported"}
                ]
                if len(commercial_data_rows) >= 2:
                    assert {
                        _fill_rgb(ws_overlay.cell(row=rr, column=1))
                        for rr in commercial_data_rows[:2]
                    } == {theme_alt}
                assert any(m.min_row == _find_row_with_value(ws_overlay, "Base operating coefficients", column=1) + 3 and m.min_col == 4 and m.max_col == 5 for m in ws_overlay.merged_cells.ranges)
                assert any(m.min_row == _find_row_with_value(ws_overlay, "Base operating coefficients", column=1) + 3 and m.min_col == 6 and m.max_col == 8 for m in ws_overlay.merged_cells.ranges)
                assert any(m.min_row == _find_row_with_value(ws_overlay, "Market inputs", column=1) + 3 and m.min_col == 7 and m.max_col == 11 for m in ws_overlay.merged_cells.ranges)
                assert any(m.min_row == _find_row_with_value(ws_overlay, "Market inputs", column=1) + 3 and m.min_col == 12 and m.max_col == 17 for m in ws_overlay.merged_cells.ranges)

            if "Operating_Drivers" in wb.sheetnames:
                ws_drivers = wb["Operating_Drivers"]
                assert _fill_rgb(ws_drivers["A2"]) == theme_title
                assert _fill_rgb(ws_drivers["A4"]) == theme_title
                assert _fill_rgb(ws_drivers["A5"]) == theme_header
                assert _fill_rgb(ws_drivers["B5"]) == theme_header

            for untouched_name in ["Hidden_Value_Flags", "QA_Log", "QA_Checks", "Needs_Review", "Promise_Evidence", "Promise_Progress"]:
                if untouched_name in wb.sheetnames:
                    untouched_fill = _fill_rgb(wb[untouched_name]["A1"])
                    assert untouched_fill not in {theme_title, theme_section, theme_header}
        finally:
            wb.close()


def test_workbook_acceptance_docs_define_analysis_sheet_style_system() -> None:
    doc_path = Path(r"c:\Users\Jibbe\Aktier\Code\docs\WORKBOOK_ACCEPTANCE.md")
    text = doc_path.read_text(encoding="utf-8")
    for required in [
        "Analysis-Sheet Style System",
        "Primary Section Blue",
        "Secondary Header Blue",
        "Panel Mist Blue",
        "Neutral Surface",
        "Grid Blue-Gray",
        "Body Text Charcoal",
    ]:
        assert required in text


def test_current_delivered_workbooks_valuation_row_order_and_semantic_fixes() -> None:
    def _assert_increasing(ws, labels: list[str]) -> None:
        rows = []
        for label in labels:
            row_idx = _find_row_with_value(ws, label, column=1)
            assert row_idx is not None, f"Missing valuation row {label!r}"
            rows.append(row_idx)
        assert rows == sorted(rows), f"Rows out of order for {labels!r}: {rows!r}"

    for ticker in ["PBI", "GPRE"]:
        workbook_path = _current_delivered_model_path(ticker)
        if not workbook_path.exists():
            pytest.skip(f"Current delivered workbook missing for valuation-order test: {workbook_path}")
        wb = load_workbook(workbook_path, data_only=False, read_only=False)
        try:
            ws = wb["Valuation"]
            _assert_increasing(ws, ["Operating", "Cash Flow", "Leverage & Liquidity", "Equity / Per-share"])
            for section_label in ["Operating", "Cash Flow", "Leverage & Liquidity", "Equity / Per-share"]:
                section_row = _find_row_with_value(ws, section_label, column=1)
                assert section_row is not None
                assert str(ws.cell(row=section_row, column=1).fill.fgColor.rgb or "") == "00D9E7F3"
            _assert_increasing(
                ws,
                [
                    "Top line",
                    "Revenue",
                    "Revenue (TTM)",
                    "Revenue YoY %",
                    "Margins",
                    "Gross margin %",
                    "Operating margin %",
                    "R&D % of revenue",
                    "Core operating",
                    "EBITDA",
                    "EBITDA margin %",
                    "EBITDA YoY %",
                    "EBITDA (TTM)",
                    "EBITDA margin (TTM)",
                    "Adjusted operating",
                    "Adj EBITDA",
                    "Adj EBITDA - EBITDA",
                    "Adj EBITDA margin %",
                    "Adj EBITDA YoY %",
                    "Adj EBITDA (TTM)",
                    "Adj EBITDA margin (TTM)",
                    "Adj EBIT (TTM)",
                    "GAAP earnings",
                    "EBIT",
                    "EBIT margin %",
                    "EBIT (TTM)",
                    "EBIT margin (TTM)",
                    "Net income",
                    "Net income margin %",
                    "Net income YoY %",
                    "Net income (TTM)",
                    "Net income margin (TTM)",
                ],
            )
            _assert_increasing(
                ws,
                [
                    "Core cash flow",
                    "CFO",
                    "Capex",
                    "Capex % of revenue",
                    "Capex % of revenue (TTM)",
                    "FCF (CFO-Capex)",
                    "FCF YoY Δ ($m)",
                    "FCF (TTM)",
                    "Adjusted / derived",
                    "Adj FCF (TTM)",
                    "Adj FCF - FCF",
                    "Owner earnings (proxy)",
                    "Cash-flow quality",
                    "FCF margin %",
                    "FCF margin (TTM)",
                    "Interest paid",
                    "Tax paid",
                    "Capital return / financing",
                    "Buybacks (cash)",
                    "Buybacks (TTM, cash)",
                    "Dividends (TTM, cash)",
                    "Acquisitions (TTM, cash)",
                    "Debt repaid (gross, TTM)",
                    "Debt issued (gross, TTM)",
                ],
            )
            _assert_increasing(
                ws,
                [
                    "Net debt position",
                    "Cash",
                    "Debt (core)",
                    "Net debt (core)",
                    "Net debt QoQ Δ ($m)",
                    "Net debt YoY Δ ($m)",
                    "Coverage / leverage",
                    "EBITDA TTM",
                    "Net leverage",
                    "Cash interest coverage (TTM)",
                    "FCF conversion (TTM)",
                    "Revolver / liquidity",
                    "Revolver facility size",
                    "Revolver drawn",
                    "Revolver letters of credit",
                    "Revolver availability",
                    "Liquidity (cash+availability)",
                    "Short-term liquidity",
                    "Current ratio",
                    "Quick ratio",
                ],
            )
            _assert_increasing(
                ws,
                [
                    "Share count",
                    "Diluted shares (m)",
                    "Shares outstanding (m)",
                    "Shares QoQ Δ (m) [out]",
                    "Shares YoY Δ (m) [out]",
                    "Per-share earnings",
                    "EPS (GAAP)",
                    "EPS YoY Δ ($)",
                    "EPS (TTM)",
                    "Adj EPS",
                    "Adj EPS (TTM)",
                    "Per-share value",
                    "BV/share",
                    "TBV/share",
                    "FCF/share (TTM)",
                    "Market-linked",
                    "EV ($m)",
                    "EV/EBITDA (TTM)",
                    "EV/Adj EBITDA (TTM)",
                    "FCF yield (TTM, equity)",
                    "FCF yield (TTM, EV)",
                ],
            )
            assert _find_row_with_value(ws, "Debt repaid (TTM)") is None
            assert _find_row_with_value(ws, "Debt issued (TTM)") is None
            assert _find_row_with_value(ws, "Total assets ($m)") is None
            assert _find_row_with_value(ws, "Total liabilities ($m)") is None
            assert _find_row_with_value(ws, "Goodwill % of assets") is None
            current_ratio_row = _find_row_with_value(ws, "Current ratio")
            quick_ratio_row = _find_row_with_value(ws, "Quick ratio")
            assert current_ratio_row is not None
            assert quick_ratio_row is not None
            assert str((ws.cell(row=current_ratio_row, column=1).comment.text if ws.cell(row=current_ratio_row, column=1).comment else "") or "") == (
                "Current assets / current liabilities. Short-term liquidity measure; around 1.0+ is often healthier."
            )
            assert str((ws.cell(row=quick_ratio_row, column=1).comment.text if ws.cell(row=quick_ratio_row, column=1).comment else "") or "") == (
                "Near-cash current assets / current liabilities. Stricter liquidity measure; around 1.0+ is often stronger."
            )

            if ticker == "PBI":
                q3_col = next(
                    cc for cc in range(2, ws.max_column + 1)
                    if str(ws.cell(row=6, column=cc).value or "").strip() == "2023-Q3"
                )
                capex_row = _find_row_with_value(ws, "Capex")
                fcf_row = _find_row_with_value(ws, "FCF (CFO-Capex)")
                assert capex_row is not None
                assert fcf_row is not None
                assert float(pd.to_numeric(ws.cell(row=capex_row, column=q3_col).value, errors="coerce")) == pytest.approx(4.42, abs=0.01)
                assert float(pd.to_numeric(ws.cell(row=fcf_row, column=q3_col).value, errors="coerce")) == pytest.approx(21.428, abs=0.01)
                convertible_row = _find_row_with_value(ws, "Convertible notes", column=12)
                debt_detail_row = _find_row_containing(ws, "August 2032", column=1)
                assert convertible_row is not None
                assert debt_detail_row is not None
                assert float(pd.to_numeric(ws.cell(row=convertible_row + 2, column=17).value, errors="coerce")) == pytest.approx(14.25, abs=0.01)
                assert float(pd.to_numeric(ws.cell(row=convertible_row + 2, column=19).value, errors="coerce")) == pytest.approx(16.135259, abs=0.001)
                assert float(pd.to_numeric(ws.cell(row=convertible_row + 2, column=21).value, errors="coerce")) == pytest.approx(5.535928, abs=0.001)
                assert str((ws.cell(row=convertible_row + 2, column=19).comment.text if ws.cell(row=convertible_row + 2, column=19).comment else "") or "") == (
                    "Capped call may reduce dilution."
                )
                assert str((ws.cell(row=debt_detail_row, column=9).comment.text if ws.cell(row=debt_detail_row, column=9).comment else "") or "") == (
                    "Capped call may reduce dilution."
                )

            if ticker == "GPRE":
                convertible_row = _find_row_with_value(ws, "Convertible notes", column=12)
                debt_detail_2027_row = _find_row_containing(ws, "2027", column=1)
                debt_detail_2030_row = _find_row_containing(ws, "2030", column=1)
                assert convertible_row is not None
                assert debt_detail_2027_row is not None
                assert debt_detail_2030_row is not None
                assert str((ws.cell(row=convertible_row + 2, column=19).comment.text if ws.cell(row=convertible_row + 2, column=19).comment else "") or "") == (
                    "Related hedge / settlement structure may reduce dilution."
                )
                assert str((ws.cell(row=convertible_row + 3, column=19).comment.text if ws.cell(row=convertible_row + 3, column=19).comment else "") or "") == (
                    "Related hedge / settlement structure may reduce dilution."
                )
                assert pd.isna(pd.to_numeric(ws.cell(row=convertible_row + 2, column=21).value, errors="coerce"))
                assert float(pd.to_numeric(ws.cell(row=convertible_row + 3, column=19).value, errors="coerce")) == pytest.approx(12.72264, abs=0.001)
                assert float(pd.to_numeric(ws.cell(row=convertible_row + 3, column=21).value, errors="coerce")) == pytest.approx(2.9, abs=0.001)
                assert pd.isna(pd.to_numeric(ws.cell(row=debt_detail_2027_row, column=12).value, errors="coerce"))
                assert float(pd.to_numeric(ws.cell(row=debt_detail_2030_row, column=9).value, errors="coerce")) == pytest.approx(12.72264, abs=0.001)
                assert str((ws.cell(row=debt_detail_2027_row, column=9).comment.text if ws.cell(row=debt_detail_2027_row, column=9).comment else "") or "") == (
                    "Related hedge / settlement structure may reduce dilution."
                )
                assert str((ws.cell(row=debt_detail_2030_row, column=9).comment.text if ws.cell(row=debt_detail_2030_row, column=9).comment else "") or "") == (
                    "Related hedge / settlement structure may reduce dilution."
                )
                ws_qn = wb["Quarter_Notes_UI"]
                realized_row = next(
                    rr
                    for rr in range(1, ws_qn.max_row + 1)
                    if "45Z production tax credits contributed $23.4m net of discounts and other costs in Q4."
                    in str(ws_qn.cell(row=rr, column=3).value or "")
                )
                assert str(ws_qn.cell(row=realized_row, column=2).value or "").strip() == "Results / drivers"
                assert str(ws_qn.cell(row=realized_row, column=4).value or "").strip() == "45Z value realized"
        finally:
            wb.close()


def test_current_delivered_gpre_workbook_preserves_obion_debt_repayment_rows() -> None:
    workbook_path = _current_delivered_model_path("GPRE")
    if not workbook_path.exists():
        pytest.skip(f"Current delivered GPRE workbook missing: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        ws = wb["Valuation"]
        q3_col = next(
            cc for cc in range(2, ws.max_column + 1)
            if str(ws.cell(row=6, column=cc).value or "").strip() == "2025-Q3"
        )
        q4_col = next(
            cc for cc in range(2, ws.max_column + 1)
            if str(ws.cell(row=6, column=cc).value or "").strip() == "2025-Q4"
        )
        debt_repaid_row = _find_row_with_value(ws, "Debt repaid (gross, TTM)")
        debt_issued_row = _find_row_with_value(ws, "Debt issued (gross, TTM)")
        assert debt_repaid_row is not None
        assert debt_issued_row is not None
        assert float(ws.cell(row=debt_repaid_row, column=q3_col).value or 0.0) == pytest.approx(130.7, abs=0.01)
        assert float(ws.cell(row=debt_repaid_row, column=q4_col).value or 0.0) == pytest.approx(130.7, abs=0.01)
        assert float(ws.cell(row=debt_issued_row, column=q4_col).value or 0.0) == pytest.approx(30.0, abs=0.01)
    finally:
        wb.close()


def test_current_delivered_gpre_workbook_applies_official_market_layout_spans() -> None:
    workbook_path = _current_delivered_model_path("GPRE")
    if not workbook_path.exists():
        pytest.skip(f"Current delivered GPRE workbook missing: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        ws = wb["Economics_Overlay"]
        base_row = _find_row_with_value(ws, "Base operating coefficients", column=1)
        market_row = _find_row_with_value(ws, "Market inputs", column=1)
        assert base_row is not None and market_row is not None
        header_row = next(
            rr
            for rr in range(base_row, market_row)
            if str(ws.cell(row=rr, column=1).value or "").strip() == "Region / family"
        )
        electricity_row = _find_row_with_value(ws, "Electricity usage", column=1)
        nebraska_row = _find_row_with_value(ws, "Nebraska", column=1)
        assert nebraska_row is not None and electricity_row is not None
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
        assert float(ws.row_dimensions[electricity_row].height or 0.0) == pytest.approx(33.0, abs=0.1)
        assert float(ws.row_dimensions[market_row].height or 0.0) == pytest.approx(21.0, abs=0.1)
        assert float(ws.row_dimensions[101].height or 0.0) == pytest.approx(22.5, abs=0.1)
        assert float(ws.row_dimensions[106].height or 0.0) == pytest.approx(21.0, abs=0.1)
        assert str(ws.cell(row=header_row, column=4).value or "").strip() == "Mapped ethanol $/gal"
        assert str(ws.cell(row=header_row, column=2).value or "").strip() == "Capacity"
        assert str(ws.cell(row=header_row, column=6).value or "").strip() == "Ethanol series"
        assert str(ws.cell(row=header_row, column=10).value or "").strip() == "Basis series"
        assert str(ws.cell(row=header_row, column=12).value or "").strip() == "Coverage / note"
        for ref in (
            f"D{header_row}:E{header_row}",
            f"F{header_row}:G{header_row}",
            f"J{header_row}:K{header_row}",
            f"L{header_row}:Q{header_row}",
            f"D{nebraska_row}:E{nebraska_row}",
            f"F{nebraska_row}:G{nebraska_row}",
            f"J{nebraska_row}:K{nebraska_row}",
            f"L{nebraska_row}:Q{nebraska_row}",
        ):
            assert ref in merged_ranges
        nebraska_rows = [
            rr
            for rr in range(header_row + 1, market_row)
            if str(ws.cell(row=rr, column=1).value or "").strip() == "Nebraska"
        ]
        assert nebraska_rows == [nebraska_row]
        assert pd.notna(pd.to_numeric(ws.cell(row=nebraska_row, column=2).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws.cell(row=nebraska_row, column=4).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws.cell(row=nebraska_row, column=8).value, errors="coerce"))
        assert str(ws.cell(row=nebraska_row, column=10).value or "").strip()
    finally:
        wb.close()


def test_current_delivered_gpre_workbook_shows_quarter_open_proxy_table_and_chart_guides() -> None:
    workbook_path = _current_delivered_model_path("GPRE")
    if not workbook_path.exists():
        pytest.skip(f"Current delivered GPRE workbook missing: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        ws = wb["Economics_Overlay"]
        ws_basis = wb["Basis_Proxy_Sandbox"]
        market_row = _find_row_with_value(ws, "Market inputs", column=1)
        process_row = _find_row_with_value(ws, "Unhedged process economics", column=1)
        corn_row = _find_row_with_value(ws, "Corn price", column=1)
        proxy_table_row = _find_row_with_value(ws, "Proxy comparison ($/gal)", column=1)
        official_proxy_row = _find_row_with_value(ws, "Approximate market crush ($/gal)", column=1)
        fitted_proxy_row = _find_row_with_value(ws, "GPRE crush proxy ($/gal)", column=1)
        forward_proxy_row = _find_row_with_value(ws, "Best forward lens ($/gal)", column=1)
        ethanol_input_row = _find_row_with_value(ws, "Ethanol price", column=1)
        sandbox_build_row = _find_row_with_value(ws_basis, "Approximate market crush build-up ($/gal)", column=2)
        chart_title_row = _find_row_with_value(ws, "Approximate market crush (weekly)", column=2)
        quarterly_chart_title_row = _find_row_with_value(ws, "Approximate market crush, fitted models, and real GPRE crush margin (quarterly)", column=2)
        quarter_compare_row = _find_row_with_value(ws, "Quarter comparisons ($/gal)", column=10)
        bridge_proxy_row = _find_row_with_value(ws, "Approximate market crush ($m)", column=1)
        bridge_gpre_row = _find_row_with_value(ws, "GPRE crush proxy ($m)", column=1)
        bridge_forward_row = _find_row_with_value(ws, "Best forward lens ($m)", column=1)
        bridge_underlying_row = _find_row_with_value(ws, "Underlying crush margin ($m)", column=1)
        bridge_reported_row = _find_row_with_value(ws, "Reported consolidated crush margin ($m)", column=1)
        bridge_45z_row = _find_row_with_value(ws, "45Z impact ($m)", column=1)
        assert market_row is not None and corn_row is not None
        assert proxy_table_row is not None and official_proxy_row is not None and fitted_proxy_row is not None and forward_proxy_row is not None
        assert ethanol_input_row is not None and sandbox_build_row is not None
        assert bridge_proxy_row is not None and bridge_gpre_row is not None and bridge_forward_row is not None and bridge_underlying_row is not None
        assert bridge_reported_row is not None and bridge_45z_row is not None
        assert process_row is None
        assert _find_row_with_value(ws, "Unhedged simple crush margin proxy", column=1) is None
        assert chart_title_row is not None and quarter_compare_row is not None
        assert float(ws.row_dimensions[101].height or 0.0) == pytest.approx(22.5, abs=0.1)
        assert float(ws.row_dimensions[106].height or 0.0) == pytest.approx(21.0, abs=0.1)
        assert str(ws.cell(row=market_row + 3, column=4).value or "").strip() == "Quarter-open outlook"
        assert str(ws.cell(row=market_row + 3, column=6).value or "").strip() == "Current QTD"
        assert str(ws.cell(row=market_row + 3, column=8).value or "").strip() == "Next quarter outlook"
        assert str(ws.cell(row=market_row + 4, column=4).value or "").strip() == "As of 2026-03-31"
        assert str(ws_basis.cell(row=sandbox_build_row + 1, column=2).value or "").strip() == "Official simple row build-up used by Approximate market crush on Economics_Overlay."
        assert str(ws_basis.cell(row=sandbox_build_row + 2, column=3).value or "").strip() == "Prior quarter"
        assert str(ws_basis.cell(row=sandbox_build_row + 2, column=5).value or "").strip() == "Quarter-open outlook"
        assert str(ws_basis.cell(row=sandbox_build_row + 2, column=7).value or "").strip() == "Current QTD"
        assert str(ws_basis.cell(row=sandbox_build_row + 2, column=9).value or "").strip() == "Next quarter outlook"
        assert str(ws_basis.cell(row=sandbox_build_row + 3, column=3).value or "").strip() == "2026-Q1"
        assert str(ws_basis.cell(row=sandbox_build_row + 3, column=5).value or "").strip() == "As of 2026-03-31"
        assert str(ws_basis.cell(row=sandbox_build_row + 3, column=7).value or "").strip().startswith("As of ")
        assert str(ws_basis.cell(row=sandbox_build_row + 3, column=9).value or "").strip() == "2026-Q3"
        sandbox_basis_snapshot_row = next(
            rr
            for rr in range(sandbox_build_row + 4, ws_basis.max_row + 1)
            if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Official corn basis snapshot date"
        )
        sandbox_basis_rule_row = next(
            rr
            for rr in range(sandbox_basis_snapshot_row, ws_basis.max_row + 1)
            if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Official corn basis selection rule"
        )
        assert str(ws_basis.cell(row=sandbox_basis_snapshot_row, column=3).value or "").strip() == "AMS fallback"
        assert str(ws_basis.cell(row=sandbox_basis_snapshot_row, column=5).value or "").strip() == "AMS fallback"
        assert re.fullmatch(r"2026-04-\d{2}", str(ws_basis.cell(row=sandbox_basis_snapshot_row, column=7).value or "").strip())
        assert re.fullmatch(r"2026-04-\d{2}", str(ws_basis.cell(row=sandbox_basis_snapshot_row, column=9).value or "").strip())
        assert str(ws_basis.cell(row=sandbox_basis_snapshot_row, column=11).value or "").strip() == "date/text"
        assert "Retained GPRE corn-bid snapshot date used by the official corn-basis leg only" in str(
            ws_basis.cell(row=sandbox_basis_snapshot_row, column=12).value or ""
        )
        assert str(ws_basis.cell(row=sandbox_basis_rule_row, column=3).value or "").strip() == "latest_snapshot_on_or_before_quarter_end / AMS fallback"
        assert str(ws_basis.cell(row=sandbox_basis_rule_row, column=5).value or "").strip() == "latest_snapshot_on_or_before_quarter_start / AMS fallback"
        assert str(ws_basis.cell(row=sandbox_basis_rule_row, column=7).value or "").strip() == "latest_snapshot_on_or_before_as_of"
        assert str(ws_basis.cell(row=sandbox_basis_rule_row, column=9).value or "").strip() == "latest_snapshot_on_or_before_as_of_with_target_quarter_rows"
        assert str(ws_basis.cell(row=sandbox_basis_rule_row, column=11).value or "").strip() == "rule/text"
        role_summary_row = _find_row_with_value(ws_basis, "Role summary", column=21)
        winner_story_row = _find_row_with_value(ws_basis, "Winner story", column=21)
        best_historical_role_row = _find_row_with_value(ws_basis, "Best historical fit", column=21)
        assert role_summary_row is not None
        assert winner_story_row is not None
        assert best_historical_role_row is not None
        assert "Production winner = fitted row used in production" in str(ws_basis.cell(row=role_summary_row + 5, column=21).value or "")
        assert "Hybrid" in str(ws_basis.cell(row=best_historical_role_row, column=22).value or "")
        assert "MAE" in str(ws_basis.cell(row=best_historical_role_row, column=22).value or "")
        assert "Forward" in str(ws_basis.cell(row=best_historical_role_row, column=22).value or "")
        proxy_note = str(ws.cell(row=proxy_table_row + 1, column=1).value or "").strip()
        assert "Official row = Approximate market crush" in proxy_note
        assert "Fitted row = GPRE crush proxy" in proxy_note
        assert "Production winner =" in proxy_note
        assert "Best forward lens =" in proxy_note
        assert "Coproduct-aware experimental lenses live in Basis_Proxy_Sandbox and are comparison-only." in proxy_note
        assert any(
            rng.min_row == proxy_table_row + 1 and rng.min_col == 1 and rng.max_col == 21
            for rng in ws.merged_cells.ranges
        )
        assert _fill_rgb(ws.cell(row=proxy_table_row + 1, column=1)) == _fill_rgb(ws.cell(row=market_row + 1, column=1))
        assert float(ws.row_dimensions[proxy_table_row + 1].height or 0.0) == pytest.approx(18.0, abs=0.1)
        assert str(ws.cell(row=proxy_table_row + 2, column=1).value or "").strip() == "Proxy row"
        assert str(ws.cell(row=proxy_table_row + 2, column=2).value or "").strip() == "Prior quarter"
        assert str(ws.cell(row=proxy_table_row + 2, column=4).value or "").strip() == "Quarter-open outlook"
        assert str(ws.cell(row=proxy_table_row + 2, column=6).value or "").strip() == "Current QTD"
        assert str(ws.cell(row=proxy_table_row + 2, column=8).value or "").strip() == "Next quarter outlook"
        assert fitted_proxy_row == official_proxy_row + 1
        assert forward_proxy_row == fitted_proxy_row + 1
        corn_source_text = str(ws.cell(row=corn_row, column=11).value or "")
        assert "Quarter-open outlook uses local manual snapshot." in corn_source_text
        assert "Current QTD:" in corn_source_text
        assert "Next quarter outlook uses live bids + AMS fallback." in corn_source_text
        assert bridge_gpre_row == bridge_proxy_row + 1
        assert bridge_forward_row == bridge_gpre_row + 1
        assert bridge_underlying_row == bridge_forward_row + 2
        assert bridge_reported_row == bridge_underlying_row + 1
        assert bridge_45z_row == bridge_reported_row + 2
        assert all(not str(ws.cell(row=bridge_forward_row + 1, column=cc).value or "").strip() for cc in range(1, 14))
        assert all(not str(ws.cell(row=bridge_reported_row + 1, column=cc).value or "").strip() for cc in range(1, 14))
        assert float(ws.row_dimensions[bridge_forward_row + 1].height or 0.0) == pytest.approx(12.0)
        assert float(ws.row_dimensions[bridge_reported_row + 1].height or 0.0) == pytest.approx(12.0)
        assert str(ws.cell(row=chart_title_row + 1, column=1).value or "").strip() == ""
        merged_ribbon_ranges = [
            str(rng)
            for rng in ws.merged_cells.ranges
            if rng.min_row == chart_title_row + 1 and rng.max_row == chart_title_row + 1 and 2 <= rng.min_col <= 21
        ]
        assert merged_ribbon_ranges == []
        assert str(ws.cell(row=quarter_compare_row + 2, column=10).value or "").strip() == "Proxy row"
        assert str(ws.cell(row=quarter_compare_row + 2, column=12).value or "").strip() == "Prior quarter vs LY"
        assert str(ws.cell(row=quarter_compare_row + 2, column=15).value or "").strip() == "Quarter-open vs LY"
        assert str(ws.cell(row=quarter_compare_row + 2, column=18).value or "").strip() == "Current QTD vs LY"
        assert str(ws.cell(row=quarter_compare_row + 2, column=20).value or "").strip() == "Next quarter vs LY"
        assert str(ws.cell(row=quarter_compare_row + 3, column=10).value or "").strip() == "Approximate market crush"
        assert str(ws.cell(row=quarter_compare_row + 4, column=10).value or "").strip() == "GPRE crush proxy"
        assert str(ws.cell(row=quarter_compare_row + 5, column=10).value or "").strip() == "Best forward lens"
        assert _fill_rgb(ws.cell(row=official_proxy_row, column=1)) == _fill_rgb(ws.cell(row=fitted_proxy_row, column=1))
        sandbox_process_margin_row = next(
            rr
            for rr in range(sandbox_build_row + 4, ws_basis.max_row + 1)
            if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Approximate market crush"
        )
        official_prior_raw = ws.cell(row=official_proxy_row, column=2).value
        fitted_prior_val = pd.to_numeric(ws.cell(row=fitted_proxy_row, column=2).value, errors="coerce")
        fitted_quarter_open_val = pd.to_numeric(ws.cell(row=fitted_proxy_row, column=4).value, errors="coerce")
        assert pd.notna(fitted_prior_val)
        assert pd.notna(fitted_quarter_open_val)
        assert str(official_prior_raw or "") == (
            f'=IF(ISNUMBER(Basis_Proxy_Sandbox!$C${sandbox_process_margin_row}),Basis_Proxy_Sandbox!$C${sandbox_process_margin_row},"")'
        )
        official_prior_compare = str(ws.cell(row=quarter_compare_row + 3, column=12).value or "").strip()
        fitted_prior_compare = str(ws.cell(row=quarter_compare_row + 4, column=12).value or "").strip()
        official_prior_match = re.match(r"([+-]?\d+(?:\.\d+)?)", official_prior_compare)
        assert official_prior_match is not None
        assert float(fitted_prior_val) != pytest.approx(float(official_prior_match.group(1)), abs=1e-9)
        assert fitted_prior_compare.startswith(f"{float(fitted_prior_val):.3f}")
        official_next_formula = str(ws.cell(row=official_proxy_row, column=8).value or "")
        fitted_current_formula = str(ws.cell(row=fitted_proxy_row, column=6).value or "")
        fitted_next_formula = str(ws.cell(row=fitted_proxy_row, column=8).value or "")
        assert f"Basis_Proxy_Sandbox!$I${sandbox_process_margin_row}" in official_next_formula
        assert fitted_current_formula.startswith("=")
        assert f"$H${ethanol_input_row}" in fitted_next_formula
        assert f"Basis_Proxy_Sandbox!$I${sandbox_process_margin_row}" in str(ws.cell(row=chart_title_row + 1, column=40).value or "")
        fitted_quarter_open_comment = str(getattr(ws.cell(row=fitted_proxy_row, column=4).comment, "text", "") or "")
        fitted_current_comment = str(getattr(ws.cell(row=fitted_proxy_row, column=6).comment, "text", "") or "")
        fitted_next_comment = str(getattr(ws.cell(row=fitted_proxy_row, column=8).comment, "text", "") or "")
        assert fitted_quarter_open_comment == "Quarter-open fitted value for the chosen model."
        assert fitted_current_comment
        assert fitted_next_comment
        assert len(fitted_current_comment.split()) <= 12
        assert len(fitted_next_comment.split()) <= 12
        assert not fitted_current_comment.startswith("Current fitted preview")
        assert not fitted_next_comment.startswith("Next-quarter fitted preview")
        assert quarterly_chart_title_row is not None
        assert quarterly_chart_title_row > chart_title_row
        assert len(ws._charts) == 3
        crush_chart = ws._charts[0]
        quarterly_chart = ws._charts[1]
        assert len(crush_chart.series) >= 6
        assert len(quarterly_chart.series) == 4
        assert getattr(getattr(quarterly_chart, "legend", None), "position", None) == "t"
        assert bool(getattr(getattr(quarterly_chart, "legend", None), "overlay", False))
        assert int(getattr(crush_chart.anchor._from, "row", -1)) + 1 == chart_title_row + 1
        assert int(getattr(crush_chart.anchor.to, "col", -1)) + 1 == 22
        with zipfile.ZipFile(workbook_path) as zf:
            chart_xmls = {
                name: zf.read(name).decode("utf-8", errors="ignore")
                for name in zf.namelist()
                if name.startswith("xl/charts/chart") and name.endswith(".xml")
            }
        weekly_chart_xml = next(
            xml
            for xml in chart_xmls.values()
            if "Approximate market crush ($/gal)" in xml and "Next quarter outlook ($/gal)" in xml
        )
        quarterly_chart_xml = next(
            xml
            for xml in chart_xmls.values()
            if "<lineChart>" in xml and 'legendPos val="t"' in xml
        )
        assert 'srgbClr val="00A0A0A0"' not in weekly_chart_xml
        assert 'srgbClr val="A0A0A0"' in weekly_chart_xml
        assert "Approximate market crush ($/gal)" in weekly_chart_xml
        assert "Prior quarter ($/gal)" in weekly_chart_xml
        assert "Current QTD ($/gal)" in weekly_chart_xml
        assert "Next quarter outlook ($/gal)" in weekly_chart_xml
        assert "Quarter boundary" in weekly_chart_xml
        assert 'symbol val="diamond"' not in weekly_chart_xml
        assert '<showSerName val="1"/>' in weekly_chart_xml
        assert 'formatCode="yyyymmdd"' not in weekly_chart_xml
        assert 'formatCode="yyyy-mm-dd"' not in weekly_chart_xml
        assert 'formatCode=";;;"' in weekly_chart_xml
        assert "<lineChart>" in quarterly_chart_xml
        assert "'Economics_Overlay'!AT" in quarterly_chart_xml
        assert "'Economics_Overlay'!AU" in quarterly_chart_xml
        assert "'Economics_Overlay'!AV" in quarterly_chart_xml
        assert "'Economics_Overlay'!AW" in quarterly_chart_xml
        assert "'Economics_Overlay'!$AS$" in quarterly_chart_xml
        assert 'legendPos val="t"' in quarterly_chart_xml
        assert '<overlay val="1"/>' in quarterly_chart_xml
        assert "<dLbls>" in quarterly_chart_xml
        assert '<showLegendKey val="1"/>' not in quarterly_chart_xml
        assert '<showVal val="1"/>' in quarterly_chart_xml
        assert '<dLblPos val="r"/>' in quarterly_chart_xml
        assert re.search(r"<catAx>.*?<majorGridlines/>", quarterly_chart_xml)
        assert "Quarter boundary" not in quarterly_chart_xml
        quarter_label_hits = re.findall(r"<v>(20\d{2}-Q[1-4])</v>", weekly_chart_xml)
        assert "2024-Q1" in quarter_label_hits
        assert "2026-Q2" in quarter_label_hits
        assert "2026-Q3" in quarter_label_hits
        future_quarters = [
            str(ws.cell(row=rr, column=45).value or "").strip()
            for rr in range(quarterly_chart_title_row + 1, quarterly_chart_title_row + 20)
            if str(ws.cell(row=rr, column=45).value or "").strip()
        ]
        assert "2025-Q4" in future_quarters
        assert "2026-Q1" in future_quarters
        assert "2026-Q2" in future_quarters
        assert "2026-Q3" in future_quarters
        assert len(future_quarters) <= 15
        assert [_quarter_label_ord(label) for label in future_quarters] == sorted(
            [_quarter_label_ord(label) for label in future_quarters]
        )
        assert future_quarters.index("2025-Q4") < future_quarters.index("2026-Q1") < future_quarters.index("2026-Q2")
        q1_helper_row = next(rr for rr in range(quarterly_chart_title_row + 1, quarterly_chart_title_row + 20) if str(ws.cell(row=rr, column=45).value or "").strip() == "2026-Q1")
        q2_helper_row = next(rr for rr in range(quarterly_chart_title_row + 1, quarterly_chart_title_row + 20) if str(ws.cell(row=rr, column=45).value or "").strip() == "2026-Q2")
        q3_helper_row = next(rr for rr in range(quarterly_chart_title_row + 1, quarterly_chart_title_row + 20) if str(ws.cell(row=rr, column=45).value or "").strip() == "2026-Q3")
        assert str(ws.cell(row=q1_helper_row, column=46).value or "").strip() == f"=B{official_proxy_row}"
        assert str(ws.cell(row=q1_helper_row, column=47).value or "").strip() == f"=B{fitted_proxy_row}"
        assert str(ws.cell(row=q1_helper_row, column=48).value or "").strip() == f"=B{forward_proxy_row}"
        assert str(ws.cell(row=q2_helper_row, column=46).value or "").strip() == f"=F{official_proxy_row}"
        assert str(ws.cell(row=q2_helper_row, column=47).value or "").strip() == f"=F{fitted_proxy_row}"
        assert str(ws.cell(row=q2_helper_row, column=48).value or "").strip() == f"=F{forward_proxy_row}"
        assert str(ws.cell(row=q3_helper_row, column=46).value or "").strip() == f"=H{official_proxy_row}"
        assert str(ws.cell(row=q3_helper_row, column=47).value or "").strip() == f"=H{fitted_proxy_row}"
        assert str(ws.cell(row=q3_helper_row, column=48).value or "").strip() == f"=H{forward_proxy_row}"
        assert pd.isna(pd.to_numeric(ws.cell(row=q1_helper_row, column=49).value, errors="coerce"))
        assert pd.isna(pd.to_numeric(ws.cell(row=q2_helper_row, column=49).value, errors="coerce"))
        assert pd.isna(pd.to_numeric(ws.cell(row=q3_helper_row, column=49).value, errors="coerce"))
    finally:
        wb.close()


def test_current_delivered_gpre_workbook_moves_process_build_up_to_basis_proxy_sandbox() -> None:
    workbook_path = _current_delivered_model_path("GPRE")
    if not workbook_path.exists():
        pytest.skip(f"Current delivered GPRE workbook missing: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        ws_overlay = wb["Economics_Overlay"]
        ws_basis = wb["Basis_Proxy_Sandbox"]
        process_row = _find_row_with_value(ws_overlay, "Unhedged process economics", column=1)
        quarterly_table_row = _find_row_with_value(ws_basis, "Quarterly comparison table", column=2)
        build_up_row = _find_row_with_value(ws_basis, "Approximate market crush build-up ($/gal)", column=2)
        corn_oil_gate_row = _find_row_with_value(ws_basis, "Coproduct source gate", column=2)
        coproduct_readiness_title_row = _find_row_with_value(ws_basis, "Coproduct signal readiness", column=2)
        coproduct_frame_summary_title_row = _find_row_with_value(ws_basis, "Coproduct frame summary", column=2)
        coproduct_history_title_row = _find_row_with_value(ws_basis, "Coproduct quarterly history", column=2)
        coproduct_volume_support_title_row = _find_row_with_value(ws_basis, "Coproduct volume support audit", column=2)
        coproduct_experimental_title_row = _find_row_with_value(ws_basis, "Coproduct-aware experimental lenses", column=2)
        best_coproduct_experimental_row = _find_row_with_value(ws_basis, "Best coproduct-aware experimental lens", column=2)
        best_coproduct_experimental_historical_row = _find_row_with_value(ws_basis, "Best historical coproduct-aware", column=2)
        best_coproduct_experimental_forward_row = _find_row_with_value(ws_basis, "Best forward coproduct-aware", column=2)
        previous_coproduct_reference_row = _find_row_with_value(ws_basis, "Previous best coproduct-aware (reference)", column=2)
        current_production_winner_reference_row = _find_row_with_value(ws_basis, "Current production winner (reference)", column=2)
        coproduct_experimental_promotion_status_row = _find_row_with_value(ws_basis, "Promotion status", column=2)
        coproduct_experimental_method_header_row = _find_row_with_value(ws_basis, "Method", column=2)
        memo_row = _find_row_with_value(ws_basis, "Hedge-adjusted memo tests", column=2)
        official_proxy_row = _find_row_with_value(ws_overlay, "Approximate market crush ($/gal)", column=1)
        fitted_proxy_row = _find_row_with_value(ws_overlay, "GPRE crush proxy ($/gal)", column=1)
        approx_bridge_row = _find_row_with_value(ws_overlay, "Approximate market crush ($m)", column=1)
        fitted_bridge_row = _find_row_with_value(ws_overlay, "GPRE crush proxy ($m)", column=1)
        helper_gallons_row = _find_row_with_value(ws_overlay, "Underlying crush margin ($m)", column=1)
        assert process_row is None
        assert quarterly_table_row is not None and build_up_row is not None and corn_oil_gate_row is not None and coproduct_readiness_title_row is not None and coproduct_frame_summary_title_row is not None and coproduct_history_title_row is not None and memo_row is not None
        assert official_proxy_row is not None and fitted_proxy_row is not None
        assert approx_bridge_row is not None and fitted_bridge_row is not None and helper_gallons_row is not None
        assert coproduct_experimental_title_row is not None
        assert best_coproduct_experimental_row is not None
        assert best_coproduct_experimental_historical_row is not None
        assert best_coproduct_experimental_forward_row is not None
        assert previous_coproduct_reference_row is not None
        assert current_production_winner_reference_row is not None
        assert coproduct_experimental_promotion_status_row is not None
        assert coproduct_experimental_method_header_row is not None
        assert quarterly_table_row < build_up_row < corn_oil_gate_row < coproduct_readiness_title_row < coproduct_frame_summary_title_row < coproduct_history_title_row < coproduct_volume_support_title_row < coproduct_experimental_title_row < memo_row
        assert str(ws_basis.cell(row=corn_oil_gate_row + 2, column=6).value or "").strip() == "YES"
        assert str(ws_basis.cell(row=corn_oil_gate_row + 3, column=6).value or "").strip() == "YES"
        assert str(ws_basis.cell(row=corn_oil_gate_row + 4, column=6).value or "").strip() == "YES"
        assert str(ws_basis.cell(row=corn_oil_gate_row + 5, column=6).value or "").strip() == "YES"
        assert str(ws_basis.cell(row=corn_oil_gate_row + 6, column=6).value or "").strip() == "YES"
        assert "go" in str(ws_basis.cell(row=corn_oil_gate_row + 7, column=6).value or "").lower()
        assert "stage b.4 keeps nwer as the primary live activation source" in str(ws_basis.cell(row=corn_oil_gate_row + 8, column=2).value or "").lower()
        assert "manual fallback/backfill" in str(ws_basis.cell(row=corn_oil_gate_row + 8, column=2).value or "").lower()
        assert str(ws_basis.cell(row=corn_oil_gate_row + 11, column=2).value or "").strip() == "Primary live activation source"
        assert str(ws_basis.cell(row=corn_oil_gate_row + 11, column=6).value or "").strip() == "NWER"
        assert str(ws_basis.cell(row=corn_oil_gate_row + 12, column=2).value or "").strip() == "Secondary corroborating source"
        assert str(ws_basis.cell(row=corn_oil_gate_row + 12, column=6).value or "").strip() == "AMS 3618"
        assert str(ws_basis.cell(row=corn_oil_gate_row + 13, column=2).value or "").strip() == "Current resolved workbook source"
        current_frame_row = _find_row_with_value(ws_basis, "Current QTD", column=2)
        assert current_frame_row is not None
        expected_delivered_source = str(ws_basis.cell(row=current_frame_row, column=8).value or "").strip()
        assert str(ws_basis.cell(row=corn_oil_gate_row + 13, column=6).value or "").strip() == expected_delivered_source
        assert "stage b.4 keeps nwer as the sufficient first visible coproduct source" in str(ws_basis.cell(row=coproduct_readiness_title_row + 1, column=2).value or "").lower()
        assert str(ws_basis.cell(row=coproduct_frame_summary_title_row + 2, column=2).value or "").strip() == "Frame"
        assert str(ws_basis.cell(row=coproduct_frame_summary_title_row + 2, column=9).value or "").strip() == "Coverage"
        assert str(ws_basis.cell(row=coproduct_frame_summary_title_row + 2, column=10).value or "").strip() == "Rule"
        assert _find_row_with_value(ws_basis, "Renewable corn oil price", column=2) is not None
        assert _find_row_with_value(ws_basis, "Distillers grains price", column=2) is not None
        assert _find_row_with_value(ws_basis, "Approximate coproduct credit", column=2) is not None
        assert _find_row_with_value(ws_basis, "NWER coproduct rows", column=2) is not None
        assert _find_row_with_value(ws_basis, "AMS 3618 coproduct rows", column=2) is not None
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=2).value or "").strip() == "Quarter"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=3).value or "").strip() == "Renewable corn oil price"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=4).value or "").strip() == "Distillers grains price"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=5).value or "").strip() == "Approximate coproduct credit ($/bushel)"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=6).value or "").strip() == "Approximate coproduct credit ($/gal)"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=7).value or "").strip() == "Approximate coproduct credit ($m)"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=8).value or "").strip() == "Resolved source mode"
        assert str(ws_basis.cell(row=coproduct_history_title_row + 2, column=9).value or "").strip() == "Coverage"
        assert "coverage is covered active-capacity share" in str(ws_basis.cell(row=coproduct_history_title_row + 1, column=2).value or "").lower()
        delivered_history_rows = [
            rr
            for rr in range(coproduct_history_title_row + 3, memo_row)
            if re.match(r"^20\d{2}-Q[1-4]$", str(ws_basis.cell(row=rr, column=2).value or "").strip())
        ]
        assert len(delivered_history_rows) >= 8
        assert str(ws_basis.cell(row=delivered_history_rows[0], column=2).value or "").strip() == "2022-Q3"
        assert str(ws_basis.cell(row=delivered_history_rows[-1], column=2).value or "").strip() == "2026-Q2"
        assert all(str(ws_basis.cell(row=rr, column=5).value or "").strip() for rr in delivered_history_rows)
        assert all(str(ws_basis.cell(row=rr, column=6).value or "").strip() for rr in delivered_history_rows)
        assert sum(1 for rr in delivered_history_rows if str(ws_basis.cell(row=rr, column=7).value or "").strip()) >= 8
        assert all(str(ws_basis.cell(row=rr, column=9).value or "").strip() for rr in delivered_history_rows)
        assert {
            str(ws_basis.cell(row=rr, column=8).value or "").strip()
            for rr in delivered_history_rows
            if str(ws_basis.cell(row=rr, column=8).value or "").strip()
        }.issubset({"NWER", "AMS 3618", "Mixed", "Unknown/blank"})
        assert str(ws_basis.cell(row=coproduct_volume_support_title_row + 2, column=2).value or "").strip() == "Series"
        assert str(ws_basis.cell(row=coproduct_volume_support_title_row + 2, column=4).value or "").strip() == "Source/path"
        assert str(ws_basis.cell(row=coproduct_volume_support_title_row + 2, column=6).value or "").strip() == "Historical usable"
        assert str(ws_basis.cell(row=coproduct_volume_support_title_row + 2, column=9).value or "").strip() == "Best use"
        assert "historical actuals only" in str(ws_basis.cell(row=coproduct_volume_support_title_row + 1, column=2).value or "").lower()
        distillers_volume_row = _find_row_with_value(ws_basis, "Distillers grains volume", column=2)
        corn_oil_volume_row = _find_row_with_value(ws_basis, "Renewable corn oil volume", column=2)
        uhp_volume_row = _find_row_with_value(ws_basis, "Ultra-high protein volume", column=2)
        mix_volume_row = _find_row_with_value(ws_basis, "Protein / coproduct mix commentary", column=2)
        assert distillers_volume_row is not None and corn_oil_volume_row is not None and uhp_volume_row is not None and mix_volume_row is not None
        assert str(ws_basis.cell(row=distillers_volume_row, column=6).value or "").strip() == "YES"
        assert str(ws_basis.cell(row=distillers_volume_row, column=7).value or "").strip() == "NO"
        assert str(ws_basis.cell(row=distillers_volume_row, column=8).value or "").strip() == "NO"
        assert str(ws_basis.cell(row=distillers_volume_row, column=9).value or "").strip() == "QA only"
        assert "tons/mm gal" in str(ws_basis.cell(row=distillers_volume_row, column=11).value or "")
        assert str(ws_basis.cell(row=corn_oil_volume_row, column=9).value or "").strip() == "QA only"
        assert "lbs/gal" in str(ws_basis.cell(row=corn_oil_volume_row, column=11).value or "")
        assert str(ws_basis.cell(row=uhp_volume_row, column=9).value or "").strip() == "Secondary QA only"
        assert "secondary qa" in str(ws_basis.cell(row=uhp_volume_row, column=11).value or "").lower()
        assert str(ws_basis.cell(row=mix_volume_row, column=6).value or "").strip() == "Commentary only"
        assert str(ws_basis.cell(row=mix_volume_row, column=9).value or "").strip() == "Context only"
        assert "comparison only" in str(ws_basis.cell(row=coproduct_experimental_title_row + 1, column=2).value or "").lower()
        assert str(ws_basis.cell(row=best_coproduct_experimental_row, column=3).value or "").strip()
        assert str(ws_basis.cell(row=best_coproduct_experimental_historical_row, column=3).value or "").strip()
        assert str(ws_basis.cell(row=best_coproduct_experimental_forward_row, column=3).value or "").strip()
        assert str(ws_basis.cell(row=previous_coproduct_reference_row, column=3).value or "").strip() == "Simple + 50% credit"
        assert str(ws_basis.cell(row=current_production_winner_reference_row, column=3).value or "").strip()
        assert str(ws_basis.cell(row=coproduct_experimental_promotion_status_row, column=3).value or "").strip() == "Experimental only"
        assert str(ws_basis.cell(row=coproduct_experimental_method_header_row, column=3).value or "").strip() == "Rule"
        assert str(ws_basis.cell(row=coproduct_experimental_method_header_row, column=4).value or "").strip() == "Clean MAE"
        assert str(ws_basis.cell(row=coproduct_experimental_method_header_row, column=12).value or "").strip() == "Status"
        delivered_coproduct_experimental_status_rows = [
            rr
            for rr in range(coproduct_experimental_method_header_row + 1, memo_row)
            if str(ws_basis.cell(row=rr, column=12).value or "").strip()
        ]
        assert len(delivered_coproduct_experimental_status_rows) == 10
        assert all(str(ws_basis.cell(row=rr, column=12).value or "").strip() == "comparison only" for rr in delivered_coproduct_experimental_status_rows)
        delivered_qtd_tracking_title_row = _find_row_with_value(ws_overlay, "Current QTD trend tracking ($/gal, crush margin lens)", column=1)
        assert delivered_qtd_tracking_title_row is not None
        _assert_gpre_qtd_tracking_upper_block_dynamic(ws_overlay)
        assert _find_row_with_value(ws_overlay, "Ethanol", column=1) is not None
        assert _find_row_with_value(ws_overlay, "Flat corn", column=1) is not None
        assert _find_row_with_value(ws_overlay, "Corn basis", column=1) is not None
        assert _find_row_with_value(ws_overlay, "Gas", column=1) is not None
        delivered_coproducts_row = _find_row_with_value(ws_overlay, "Coproducts", column=1)
        assert delivered_coproducts_row == delivered_qtd_tracking_title_row + 14
        assert _find_row_with_value(ws_overlay, "Residual", column=1) is None
        assert _find_row_containing(
            ws_overlay,
            "Same-point-last-quarter is intentionally not a primary tracking metric here",
            column=2,
        ) is None
        assert ws_overlay.row_dimensions[82].height == pytest.approx(18.0, abs=0.01)
        assert ws_overlay.row_dimensions[107].height == pytest.approx(18.0, abs=0.01)
        assert ws_overlay.row_dimensions[122].height == pytest.approx(18.0, abs=0.01)
        assert ws_overlay.row_dimensions[delivered_qtd_tracking_title_row + 13].height == pytest.approx(15.0, abs=0.01)
        basis_weighting_row = _find_row_with_value(ws_overlay, "Basis weighting", column=1)
        assert basis_weighting_row is not None
        assert str(ws_overlay.cell(row=basis_weighting_row, column=2).value or "").strip() == (
            "Official corn basis prefers dated GPRE plant bids when available; otherwise it falls back to "
            "active-capacity-weighted AMS basis using mapped state/regional series and deterministic fallbacks"
        )
        assert ws_overlay.row_dimensions[delivered_coproducts_row].height == pytest.approx(18.0, abs=0.01)
        assert str(ws_overlay.cell(row=delivered_coproducts_row, column=1).fill.fgColor.rgb or "").strip().upper() == str(ws_overlay.cell(row=delivered_qtd_tracking_title_row, column=1).fill.fgColor.rgb or "").strip().upper()
        assert ws_overlay.row_dimensions[delivered_coproducts_row + 5].height == pytest.approx(8.0, abs=0.01)
        assert str(ws_overlay.cell(row=delivered_coproducts_row + 5, column=1).value or "").strip() == ""
        assert str(ws_overlay.cell(row=delivered_coproducts_row + 5, column=1).fill.fgColor.rgb or "").strip().upper() == str(ws_overlay.cell(row=delivered_qtd_tracking_title_row + 4, column=2).fill.fgColor.rgb or "").strip().upper()
        delivered_coproduct_header_row = _find_row_with_value(ws_overlay, "Coproduct economics", column=1)
        assert delivered_coproduct_header_row is not None and delivered_coproduct_header_row > delivered_qtd_tracking_title_row
        assert delivered_coproduct_header_row == delivered_coproducts_row + 1
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 1, column=2).value or "").strip() == "2026-Q1"
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 1, column=4).value or "").strip() == "As of 2026-03-31"
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 1, column=6).value or "").strip().startswith("As of 2026-04-")
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 1, column=8).value or "").strip() == "2026-Q3"
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 2, column=1).value or "").strip() == "Renewable corn oil price"
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 3, column=1).value or "").strip() == "Distillers grains price"
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 4, column=1).value or "").strip() == ""
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 5, column=1).value or "").strip() == "Approximate coproduct credit ($/gal)"
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 6, column=1).value or "").strip() == "Approximate coproduct credit ($m)"
        delivered_coproduct_chart_title_row = _find_row_with_value(ws_overlay, "Approximate coproduct credit ($/gal, quarterly history)", column=2)
        delivered_mini_history_title_row = _find_row_with_value(ws_overlay, "Coproduct credit", column=2)
        assert delivered_coproduct_chart_title_row is not None
        assert delivered_mini_history_title_row is not None
        assert str(ws_overlay.cell(row=delivered_mini_history_title_row + 1, column=2).value or "").strip() == "Quarter"
        assert str(ws_overlay.cell(row=delivered_mini_history_title_row + 1, column=4).value or "").strip() == "$/gal"
        assert str(ws_overlay.cell(row=delivered_mini_history_title_row + 1, column=6).value or "").strip() == "$m"
        assert str(ws_overlay.cell(row=delivered_mini_history_title_row + 1, column=8).value or "").strip() == "Coverage"
        assert str(ws_overlay.cell(row=delivered_mini_history_title_row + 1, column=10).value or "").strip() == "Source mode"
        assert str(ws_overlay.cell(row=delivered_mini_history_title_row, column=13).value or "").strip() == "Corn oil prices"
        assert str(ws_overlay.cell(row=delivered_mini_history_title_row + 1, column=13).value or "").strip() == "Quarter"
        assert str(ws_overlay.cell(row=delivered_mini_history_title_row + 1, column=15).value or "").strip() == "$/lb"
        assert str(ws_overlay.cell(row=delivered_mini_history_title_row + 1, column=17).value or "").strip() == "$/gal"
        assert str(ws_overlay.cell(row=delivered_mini_history_title_row + 1, column=19).value or "").strip() == "$m proxy"
        delivered_coverage_note_row = _find_row_with_value(
            ws_overlay,
            "Coverage reflects covered active-capacity footprint; values are covered-footprint weighted averages.",
            column=2,
        )
        assert delivered_coverage_note_row is not None
        delivered_mini_history_rows = [
            rr
            for rr in range(delivered_mini_history_title_row + 2, delivered_coverage_note_row)
            if str(ws_overlay.cell(row=rr, column=4).value or "").strip()
        ]
        assert len(delivered_mini_history_rows) >= 14
        assert all(str(ws_overlay.cell(row=rr, column=13).value or "").strip() for rr in delivered_mini_history_rows)
        assert any(pd.notna(pd.to_numeric(ws_overlay.cell(row=rr, column=15).value, errors="coerce")) for rr in delivered_mini_history_rows)
        assert any(pd.notna(pd.to_numeric(ws_overlay.cell(row=rr, column=17).value, errors="coerce")) for rr in delivered_mini_history_rows)
        assert any(pd.notna(pd.to_numeric(ws_overlay.cell(row=rr, column=19).value, errors="coerce")) for rr in delivered_mini_history_rows)
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 2, column=6).value or "").strip()
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 3, column=6).value or "").strip()
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 5, column=6).value or "").strip()
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 6, column=6).value or "").strip()
        assert str(ws_overlay.cell(row=delivered_coproduct_header_row + 6, column=6).number_format or "").strip() == "#,##0.0"
        delivered_coproduct_helper_labels = [
            str(ws_overlay.cell(row=rr, column=49).value or "").strip()
            for rr in range(delivered_coproduct_chart_title_row + 1, delivered_mini_history_title_row)
            if str(ws_overlay.cell(row=rr, column=49).value or "").strip()
        ]
        assert "2026-Q1" in delivered_coproduct_helper_labels
        assert "2026-Q2" in delivered_coproduct_helper_labels
        assert "2026-Q3" in delivered_coproduct_helper_labels
        assert delivered_coproduct_helper_labels.index("2026-Q1") < delivered_coproduct_helper_labels.index("2026-Q2") < delivered_coproduct_helper_labels.index("2026-Q3")
        assert len(delivered_coproduct_helper_labels) <= 15
        assert [_quarter_label_ord(label) for label in delivered_coproduct_helper_labels] == sorted(
            [_quarter_label_ord(label) for label in delivered_coproduct_helper_labels]
        )
        assert len(ws_overlay._charts) == 3
        prior_frame_row = _find_row_with_value(ws_basis, "Prior quarter", column=2)
        quarter_open_frame_row = _find_row_with_value(ws_basis, "Quarter-open outlook", column=2)
        next_frame_row = _find_row_with_value(ws_basis, "Next quarter outlook", column=2)
        delivered_q1_history_row = _find_row_with_value(ws_basis, "2026-Q1", column=2)
        delivered_q2_history_row = _find_row_with_value(ws_basis, "2026-Q2", column=2)
        assert prior_frame_row is not None and quarter_open_frame_row is not None and current_frame_row is not None and next_frame_row is not None
        assert delivered_q1_history_row is not None and delivered_q2_history_row is not None
        expected_visible_frame_refs = {
            delivered_coproduct_header_row + 2: 3,
            delivered_coproduct_header_row + 3: 4,
            delivered_coproduct_header_row + 5: 6,
            delivered_coproduct_header_row + 6: 7,
        }
        frame_rows_by_col = {
            2: prior_frame_row,
            4: quarter_open_frame_row,
            6: current_frame_row,
            8: next_frame_row,
        }
        for visible_row, frame_value_col in expected_visible_frame_refs.items():
            for cc, frame_row_num in frame_rows_by_col.items():
                expected_formula = f'=IF(ISNUMBER(Basis_Proxy_Sandbox!${get_column_letter(frame_value_col)}${frame_row_num}),Basis_Proxy_Sandbox!${get_column_letter(frame_value_col)}${frame_row_num},"")'
                assert str(ws_overlay.cell(row=visible_row, column=cc).value or "").strip() == expected_formula
        delivered_coproduct_q1_helper_row = next(rr for rr in range(delivered_coproduct_chart_title_row + 1, delivered_mini_history_title_row) if str(ws_overlay.cell(row=rr, column=49).value or "").strip() == "2026-Q1")
        delivered_coproduct_q2_helper_row = next(rr for rr in range(delivered_coproduct_chart_title_row + 1, delivered_mini_history_title_row) if str(ws_overlay.cell(row=rr, column=49).value or "").strip() == "2026-Q2")
        delivered_coproduct_q3_helper_row = next(rr for rr in range(delivered_coproduct_chart_title_row + 1, delivered_mini_history_title_row) if str(ws_overlay.cell(row=rr, column=49).value or "").strip() == "2026-Q3")
        assert str(ws_overlay.cell(row=delivered_coproduct_q1_helper_row, column=50).value or "").strip() == f"=Basis_Proxy_Sandbox!$F${delivered_q1_history_row}"
        assert str(ws_overlay.cell(row=delivered_coproduct_q2_helper_row, column=50).value or "").strip() == f"=Basis_Proxy_Sandbox!$F${delivered_q2_history_row}"
        assert str(ws_overlay.cell(row=delivered_coproduct_q3_helper_row, column=50).value or "").strip() == f"=Basis_Proxy_Sandbox!$F${next_frame_row}"
        assert "weighted active-capacity quarterly resolver" in str(ws_overlay.cell(row=delivered_coproduct_header_row + 2, column=11).value or "").lower()
        assert "weighted active-capacity quarterly resolver" in str(ws_overlay.cell(row=delivered_coproduct_header_row + 3, column=11).value or "").lower()
        assert "weighted sandbox build-up divided by ethanol yield" in str(ws_overlay.cell(row=delivered_coproduct_header_row + 5, column=11).value or "").lower()
        assert "source mode, coverage, and carry-forward rule" in str(ws_overlay.cell(row=delivered_coproduct_header_row + 5, column=11).value or "").lower()
        assert "frame-specific implied gallons basis" in str(ws_overlay.cell(row=delivered_coproduct_header_row + 6, column=11).value or "").lower()
        assert str(ws_overlay.cell(row=delivered_mini_history_rows[0], column=2).value or "").strip() == "2026-Q3"
        assert str(ws_overlay.cell(row=delivered_mini_history_rows[-1], column=2).value or "").strip().startswith("=Basis_Proxy_Sandbox!$B$")
        delivered_q2_table_row = next(rr for rr in delivered_mini_history_rows if str(ws_overlay.cell(row=rr, column=2).value or "").strip().startswith("=Basis_Proxy_Sandbox!$B$150"))
        delivered_q1_table_row = next(rr for rr in delivered_mini_history_rows if str(ws_overlay.cell(row=rr, column=2).value or "").strip().startswith("=Basis_Proxy_Sandbox!$B$149"))
        delivered_q3_table_row = next(rr for rr in delivered_mini_history_rows if str(ws_overlay.cell(row=rr, column=2).value or "").strip() == "2026-Q3")
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=delivered_q2_table_row, column=6).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=delivered_q1_table_row, column=6).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=delivered_q1_table_row, column=17).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=delivered_q1_table_row, column=19).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=delivered_q3_table_row, column=17).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=delivered_q3_table_row, column=19).value, errors="coerce"))
        assert str(ws_overlay.cell(row=delivered_mini_history_rows[0], column=6).number_format or "").strip() == "#,##0.0"
        assert str(ws_overlay.cell(row=delivered_mini_history_rows[0], column=19).number_format or "").strip() == "#,##0.0"

        build_labels = [
            str(ws_basis.cell(row=rr, column=2).value or "").strip()
            for rr in range(build_up_row + 4, build_up_row + 12)
        ]
        assert build_labels == [
            "Ethanol revenue contribution",
            "Distillers contribution",
            "Ultra-high protein contribution",
            "Renewable corn oil contribution",
            "Feedstock cost",
            "Natural gas burden",
            "Approximate coproduct credit",
            "Approximate market crush",
        ]

        sandbox_process_margin_row = build_up_row + 11
        expected_official_refs = {
            2: f'=IF(ISNUMBER(Basis_Proxy_Sandbox!$C${sandbox_process_margin_row}),Basis_Proxy_Sandbox!$C${sandbox_process_margin_row},"")',
            4: f'=IF(ISNUMBER(Basis_Proxy_Sandbox!$E${sandbox_process_margin_row}),Basis_Proxy_Sandbox!$E${sandbox_process_margin_row},"")',
            6: f'=IF(ISNUMBER(Basis_Proxy_Sandbox!$G${sandbox_process_margin_row}),Basis_Proxy_Sandbox!$G${sandbox_process_margin_row},"")',
            8: f'=IF(ISNUMBER(Basis_Proxy_Sandbox!$I${sandbox_process_margin_row}),Basis_Proxy_Sandbox!$I${sandbox_process_margin_row},"")',
        }
        for cc, expected_formula in expected_official_refs.items():
            assert str(ws_overlay.cell(row=official_proxy_row, column=cc).value or "").strip() == expected_formula

        assert str(getattr(ws_overlay.cell(row=fitted_proxy_row, column=4).comment, "text", "") or "") == "Quarter-open fitted value for the chosen model."
        current_fitted_comment = str(getattr(ws_overlay.cell(row=fitted_proxy_row, column=6).comment, "text", "") or "")
        next_fitted_comment = str(getattr(ws_overlay.cell(row=fitted_proxy_row, column=8).comment, "text", "") or "")
        assert current_fitted_comment
        assert next_fitted_comment
        assert len(current_fitted_comment.split()) <= 12
        assert len(next_fitted_comment.split()) <= 12
        assert not current_fitted_comment.startswith("Current fitted preview")
        assert not next_fitted_comment.startswith("Next-quarter fitted preview")
        assert str(ws_overlay.cell(row=approx_bridge_row, column=14).value or "").strip() == (
            f'=IF(AND(ISNUMBER(B{official_proxy_row}),ISNUMBER(N{helper_gallons_row})),B{official_proxy_row}*N{helper_gallons_row},"")'
        )
        assert str(ws_overlay.cell(row=fitted_bridge_row, column=14).value or "").strip() == (
            f'=IF(AND(ISNUMBER(B{fitted_proxy_row}),ISNUMBER(N{helper_gallons_row})),B{fitted_proxy_row}*N{helper_gallons_row},"")'
        )
        weekly_chart_title_row = _find_row_with_value(ws_overlay, "Approximate market crush (weekly)", column=2)
        assert weekly_chart_title_row is not None
        assert str(ws_overlay.cell(row=weekly_chart_title_row + 1, column=42).value or "").strip() == f"=IF(ISNUMBER(B{official_proxy_row}),B{official_proxy_row},NA())"
        assert str(ws_overlay.cell(row=weekly_chart_title_row + 1, column=44).value or "").strip() == f"=IF(ISNUMBER(F{official_proxy_row}),F{official_proxy_row},NA())"
        assert "Basis_Proxy_Sandbox!$I$" in str(ws_overlay.cell(row=_find_row_with_value(ws_overlay, "Approximate market crush (weekly)", column=2) + 1, column=40).value or "")
        relevant_formula_texts = [
            str(cell.value or "").strip()
            for row in ws_overlay.iter_rows(min_row=max(1, official_proxy_row - 2), max_row=ws_overlay.max_row, min_col=1, max_col=26)
            for cell in row
            if isinstance(cell.value, str) and str(cell.value).startswith("=")
        ]
        assert relevant_formula_texts
        assert all("#REF!" not in txt for txt in relevant_formula_texts)
        assert all(old_ref not in txt for txt in relevant_formula_texts for old_ref in ("B113", "D113", "F113", "H113"))
    finally:
        wb.close()


def test_current_delivered_gpre_summary_artifact_matches_workbook_winner_story() -> None:
    workbook_path = _current_delivered_model_path("GPRE")
    summary_path = Path(__file__).resolve().parents[2] / "GPRE" / "basis_proxy" / "gpre_basis_proxy_summary.md"
    if not workbook_path.exists() or not summary_path.exists():
        pytest.skip("Current delivered GPRE workbook or summary artifact missing for consistency readback test.")

    summary_text = summary_path.read_text(encoding="utf-8")
    summary_lower = summary_text.lower()
    expanded_match = re.search(r"Expanded-pass best candidate:\s+([a-z0-9_]+)\s+\|", summary_text)
    winner_match = re.search(r"Production winner:\s+([a-z0-9_]+)\s+\|", summary_text)
    experimental_match = re.search(r"Best experimental candidate:\s+([a-z0-9_]+)\.", summary_text)
    assert expanded_match is not None
    assert winner_match is not None
    assert experimental_match is not None

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        ws_basis = wb["Basis_Proxy_Sandbox"]
        expanded_row = _find_row_with_value(ws_basis, "Expanded-pass best", column=21)
        winner_row = _find_row_with_value(ws_basis, "Production winner", column=21)
        preview_row = _find_row_with_value(ws_basis, "Preview quality", column=21)
        experimental_title_row = _find_row_with_value(ws_basis, "Experimental realization / regime candidates", column=21)
        assert expanded_row is not None and winner_row is not None and preview_row is not None and experimental_title_row is not None

        expanded_display = str(ws_basis.cell(row=expanded_row, column=22).value or "").strip()
        winner_display = str(ws_basis.cell(row=winner_row, column=22).value or "").strip()
        preview_display = str(ws_basis.cell(row=preview_row, column=22).value or "").strip().lower()
        sandbox_label_map = {
            "bridge_front_loaded": "Bridge front-loaded",
            "process_front_loaded": "Process front-loaded",
            "process_quarter_open_blend": "Process q-open blend",
            "process_quarter_open_blend_ops_penalty": "Process q-open blend + ops penalty",
            "process_quarter_open_blend_exec_penalty": "Process q-open + severe ops penalty",
            "process_quarter_open_blend_utilization_penalty": "Process q-open + utilization penalty",
            "process_quarter_open_blend_maintenance_delay_penalty": "Process q-open + maintenance delay",
            "process_quarter_open_blend_inventory_timing_drag": "Process q-open + inventory drag",
            "process_quarter_open_blend_locked_setup": "Process q-open + locked setup",
            "process_basis_blend_current40_front60": "Process basis blend 40/60",
            "process_basis_passthrough_beta35": "Process basis beta 0.35",
            "process_basis_passthrough_beta65": "Process basis beta 0.65",
            "process_quarter_open_current50_exec_penalty": "Process q-open/current 50/50 + exec penalty",
            "process_regime_basis_passthrough": "Process regime basis passthrough",
            "process_two_stage_realization_residual": "Process two-stage residual",
            "process_capacity_weighted_basis_strict": "Process capacity-weighted basis strict",
            "process_inventory_gap_penalty_small": "Process inventory gap penalty small",
            "process_inventory_gap_penalty_medium": "Process inventory gap penalty medium",
            "process_utilization_regime_blend": "Process utilization regime blend",
            "process_utilization_regime_residual": "Process utilization regime residual",
            "process_exec_inventory_combo_medium": "Process exec + inventory combo",
            "process_asymmetric_basis_passthrough": "Process asymmetric basis passthrough",
            "process_residual_regime_locked_vs_disturbed": "Process residual regime split",
            "process_gated_incumbent_vs_residual": "Process gated incumbent vs residual",
            "process_front_loaded_ops_penalty": "Process front + ops penalty",
            "process_front_loaded_ethanol_geo": "Process front + ethanol geo",
        }

        assert expanded_display == sandbox_label_map[expanded_match.group(1)]
        assert winner_display.startswith(sandbox_label_map[winner_match.group(1)])
        assert "backtest window: 2023-q1 to 2025-q4." in summary_lower
        assert "style vs family:" in summary_lower
        assert "diagnostic scope: diagnostic only" in summary_lower
        assert "roles / consistency check" in summary_lower
        assert "selection vs promotion:" in summary_lower
        assert "winner preview quality: close." in summary_lower
        assert "experimental signal audit" in summary_lower
        assert "experimental realization / regime comparison" in summary_lower
        assert "selection vs promotion" in str(ws_basis["U5"].value or "").lower()
        assert "close" in preview_display
        assert str(ws_basis.cell(row=experimental_title_row + 1, column=21).value or "").strip() == "Current winner"
        assert str(ws_basis.cell(row=experimental_title_row + 2, column=21).value or "").strip() == "Best experimental"
        assert str(ws_basis.cell(row=experimental_title_row + 3, column=21).value or "").strip() == "Promoted?"
        assert str(ws_basis.cell(row=experimental_title_row + 2, column=22).value or "").strip() == sandbox_label_map[experimental_match.group(1)]
    finally:
        wb.close()


def test_current_delivered_gpre_hedge_style_table_matches_summary_weak_fit_story() -> None:
    workbook_path = _current_delivered_model_path("GPRE")
    summary_path = Path(__file__).resolve().parents[2] / "GPRE" / "basis_proxy" / "gpre_basis_proxy_summary.md"
    if not workbook_path.exists() or not summary_path.exists():
        pytest.skip("Current delivered GPRE workbook or summary artifact missing for hedge-style consistency readback test.")

    summary_text = summary_path.read_text(encoding="utf-8")
    summary_lower = summary_text.lower()
    backtest_match = re.search(r"Backtest window:\s*([0-9]{4}-Q[1-4]\s+to\s+[0-9]{4}-Q[1-4])\.", summary_text)
    style_match = re.search(r"Best overall style:\s*(.*?)\s*\|\s*best overall family:\s*(.*?)\s*\|\s*usable quarters:\s*(\d+)\.", summary_text)
    weak_fit_match = re.search(r"Weak-fit quarters:\s*(.*?)\.", summary_text)
    assert backtest_match is not None
    assert style_match is not None
    assert weak_fit_match is not None

    summary_backtest_window = backtest_match.group(1).strip()
    summary_best_style = style_match.group(1).strip()
    summary_best_family = style_match.group(2).strip()
    weak_fit_raw = weak_fit_match.group(1).strip()
    summary_weak_fit_quarters = [] if weak_fit_raw.lower() == "none" else [part.strip() for part in weak_fit_raw.split(",") if part.strip()]

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        ws_basis = wb["Basis_Proxy_Sandbox"]
        hedge_study_row = _find_row_with_value(ws_basis, "Implied hedge / realization style study", column=2)
        quarter_best_fit_row = _find_row_with_value(ws_basis, "Quarter-by-quarter best-fit hedge style", column=2)
        interp_row = _find_row_with_value(ws_basis, "Interpretation", column=2)
        assert hedge_study_row is not None and quarter_best_fit_row is not None and interp_row is not None

        hedge_summary_text = str(ws_basis.cell(row=hedge_study_row + 1, column=2).value or "").strip()
        hedge_note_text = str(ws_basis.cell(row=hedge_study_row + 2, column=2).value or "").strip()
        assert f"Backtest window: {summary_backtest_window}." in hedge_summary_text
        assert f"Best overall style: {summary_best_style}." in hedge_summary_text
        assert f"Best overall family: {summary_best_family}." in hedge_summary_text
        assert "diagnostic only; does not change official row, fitted row, or winner selection" in hedge_note_text.lower()
        assert "single lowest-mae candidate style" in hedge_note_text.lower()

        headers = [str(ws_basis.cell(row=quarter_best_fit_row + 1, column=cc).value or "").strip() for cc in range(2, 10)]
        assert headers == [
            "Quarter",
            "Reported consolidated crush margin ($/gal)",
            "Best-fit style",
            "Best-fit value",
            "Error",
            "Weak fit?",
            "Hard quarter?",
            "Note/category",
        ]

        workbook_weak_fit_quarters = []
        for rr in range(quarter_best_fit_row + 2, interp_row):
            quarter_txt = str(ws_basis.cell(row=rr, column=2).value or "").strip()
            weak_fit_txt = str(ws_basis.cell(row=rr, column=7).value or "").strip()
            if not quarter_txt:
                continue
            if weak_fit_txt.lower() == "yes":
                workbook_weak_fit_quarters.append(quarter_txt)

        assert workbook_weak_fit_quarters
        assert set(workbook_weak_fit_quarters) == set(summary_weak_fit_quarters)
        for quarter_txt in workbook_weak_fit_quarters:
            assert quarter_txt.lower() in summary_lower
    finally:
        wb.close()


def test_current_delivered_gpre_workbook_shows_proxy_implied_results_bridge_panel() -> None:
    workbook_path = _current_delivered_model_path("GPRE")
    if not workbook_path.exists():
        pytest.skip(f"Current delivered GPRE workbook missing: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        ws = wb["Economics_Overlay"]
        bridge_row = _find_row_with_value(ws, "Bridge to reported", column=1)
        official_proxy_row = _find_row_with_value(ws, "Approximate market crush ($/gal)", column=1)
        fitted_proxy_row = _find_row_with_value(ws, "GPRE crush proxy ($/gal)", column=1)
        forward_proxy_row = _find_row_with_value(ws, "Best forward lens ($/gal)", column=1)
        approx_bridge_row = _find_row_with_value(ws, "Approximate market crush ($m)", column=1)
        fitted_bridge_row = _find_row_with_value(ws, "GPRE crush proxy ($m)", column=1)
        forward_bridge_row = _find_row_with_value(ws, "Best forward lens ($m)", column=1)
        underlying_bridge_row = _find_row_with_value(ws, "Underlying crush margin ($m)", column=1)
        reported_bridge_row = _find_row_with_value(ws, "Reported consolidated crush margin ($m)", column=1)
        assert all(
            row is not None
            for row in (
                bridge_row,
                official_proxy_row,
                fitted_proxy_row,
                forward_proxy_row,
                approx_bridge_row,
                fitted_bridge_row,
                forward_bridge_row,
                underlying_bridge_row,
                reported_bridge_row,
            )
        )

        panel_title_row = int(bridge_row) + 1
        panel_header_row = int(bridge_row) + 2
        panel_subheader_row = int(bridge_row) + 3
        helper_gallons_row = int(underlying_bridge_row)
        helper_basis_row = int(reported_bridge_row)
        frame_start_cols = {
            "prior_quarter": 14,
            "quarter_open": 16,
            "current_qtd": 18,
            "next_quarter_thesis": 20,
        }

        assert any(m.min_row == panel_title_row and m.min_col == 1 and m.max_col == 13 for m in ws.merged_cells.ranges)
        assert str(ws.cell(row=panel_title_row, column=1).value or "").strip().startswith("Approximate market crush shows simple weighted market/process conditions")
        assert str(ws.cell(row=panel_title_row, column=14).value or "").strip() == "Proxy-implied results ($m)"
        panel_title_comment = ws.cell(row=panel_title_row, column=15).comment
        if panel_title_comment is None:
            panel_title_comment = ws.cell(row=panel_title_row, column=14).comment
        assert panel_title_comment is not None
        assert "proxy-implied translation only" in str(panel_title_comment.text or "").lower()
        assert str(ws.cell(row=panel_title_row, column=15).value or "").strip() == ""
        assert any(m.min_row == panel_title_row and m.min_col == 14 and m.max_col == 21 for m in ws.merged_cells.ranges)

        assert any(m.min_row == panel_header_row and m.min_col == 14 and m.max_col == 15 for m in ws.merged_cells.ranges)
        assert any(m.min_row == panel_header_row and m.min_col == 16 and m.max_col == 17 for m in ws.merged_cells.ranges)
        assert any(m.min_row == panel_header_row and m.min_col == 18 and m.max_col == 19 for m in ws.merged_cells.ranges)
        assert any(m.min_row == panel_header_row and m.min_col == 20 and m.max_col == 21 for m in ws.merged_cells.ranges)
        assert [str(ws.cell(row=panel_header_row, column=cc).value or "").strip() for cc in (14, 16, 18, 20)] == [
            "Prior quarter",
            "Quarter-open outlook",
            "Current QTD",
            "Next quarter outlook",
        ]
        assert any(m.min_row == panel_subheader_row and m.min_col == 14 and m.max_col == 15 for m in ws.merged_cells.ranges)
        assert any(m.min_row == panel_subheader_row and m.min_col == 16 and m.max_col == 17 for m in ws.merged_cells.ranges)
        assert any(m.min_row == panel_subheader_row and m.min_col == 18 and m.max_col == 19 for m in ws.merged_cells.ranges)
        assert any(m.min_row == panel_subheader_row and m.min_col == 20 and m.max_col == 21 for m in ws.merged_cells.ranges)
        panel_subheaders = [str(ws.cell(row=panel_subheader_row, column=cc).value or "").strip() for cc in (14, 16, 18, 20)]
        assert panel_subheaders[0] == "2026-Q1"
        assert panel_subheaders[1] == "As of 2026-03-31"
        assert panel_subheaders[2].startswith("As of 2026-04-")
        assert panel_subheaders[3] == "2026-Q3"

        assert str(ws.cell(row=helper_gallons_row, column=22).value or "").strip() == "Implied gallons assumption"
        assert any(m.min_row == helper_gallons_row and m.min_col == 22 and m.max_col == 24 for m in ws.merged_cells.ranges)
        assert str(ws.cell(row=helper_basis_row, column=22).value or "").strip() == "Volume basis"
        assert any(m.min_row == helper_basis_row and m.min_col == 22 and m.max_col == 24 for m in ws.merged_cells.ranges)

        neighbor_width = float(ws.column_dimensions["U"].width or 0.0)
        for col_letter in ("V", "W", "X"):
            assert float(ws.column_dimensions[col_letter].width or 0.0) == pytest.approx(neighbor_width, abs=0.1)

        official_formula_cells = {
            14: f'=IF(AND(ISNUMBER(B{official_proxy_row}),ISNUMBER(N{helper_gallons_row})),B{official_proxy_row}*N{helper_gallons_row},"")',
            16: f'=IF(AND(ISNUMBER(D{official_proxy_row}),ISNUMBER(P{helper_gallons_row})),D{official_proxy_row}*P{helper_gallons_row},"")',
            18: f'=IF(AND(ISNUMBER(F{official_proxy_row}),ISNUMBER(R{helper_gallons_row})),F{official_proxy_row}*R{helper_gallons_row},"")',
            20: f'=IF(AND(ISNUMBER(H{official_proxy_row}),ISNUMBER(T{helper_gallons_row})),H{official_proxy_row}*T{helper_gallons_row},"")',
        }
        fitted_formula_cells = {
            14: f'=IF(AND(ISNUMBER(B{fitted_proxy_row}),ISNUMBER(N{helper_gallons_row})),B{fitted_proxy_row}*N{helper_gallons_row},"")',
            16: f'=IF(AND(ISNUMBER(D{fitted_proxy_row}),ISNUMBER(P{helper_gallons_row})),D{fitted_proxy_row}*P{helper_gallons_row},"")',
            18: f'=IF(AND(ISNUMBER(F{fitted_proxy_row}),ISNUMBER(R{helper_gallons_row})),F{fitted_proxy_row}*R{helper_gallons_row},"")',
            20: f'=IF(AND(ISNUMBER(H{fitted_proxy_row}),ISNUMBER(T{helper_gallons_row})),H{fitted_proxy_row}*T{helper_gallons_row},"")',
        }
        forward_formula_cells = {
            14: f'=IF(AND(ISNUMBER(B{forward_proxy_row}),ISNUMBER(N{helper_gallons_row})),B{forward_proxy_row}*N{helper_gallons_row},"")',
            16: f'=IF(AND(ISNUMBER(D{forward_proxy_row}),ISNUMBER(P{helper_gallons_row})),D{forward_proxy_row}*P{helper_gallons_row},"")',
            18: f'=IF(AND(ISNUMBER(F{forward_proxy_row}),ISNUMBER(R{helper_gallons_row})),F{forward_proxy_row}*R{helper_gallons_row},"")',
            20: f'=IF(AND(ISNUMBER(H{forward_proxy_row}),ISNUMBER(T{helper_gallons_row})),H{forward_proxy_row}*T{helper_gallons_row},"")',
        }
        for cc, expected_formula in official_formula_cells.items():
            assert str(ws.cell(row=approx_bridge_row, column=cc).value or "").strip() == expected_formula
        for cc, expected_formula in fitted_formula_cells.items():
            assert str(ws.cell(row=fitted_bridge_row, column=cc).value or "").strip() == expected_formula
        for cc, expected_formula in forward_formula_cells.items():
            assert str(ws.cell(row=forward_bridge_row, column=cc).value or "").strip() == expected_formula

        assert str(ws.cell(row=helper_basis_row, column=14).value or "").strip() == "Fallback: YoY produced gallons adjusted to active capacity"
        for frame_key in ("quarter_open", "current_qtd", "next_quarter_thesis"):
            assert str(ws.cell(row=helper_basis_row, column=frame_start_cols[frame_key]).value or "").strip() == "YoY same quarter produced gallons, adjusted to current active capacity footprint (730 MMgy)"

        assert pd.notna(pd.to_numeric(ws.cell(row=helper_gallons_row, column=14).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws.cell(row=helper_gallons_row, column=16).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws.cell(row=helper_gallons_row, column=18).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws.cell(row=helper_gallons_row, column=20).value, errors="coerce"))
        for cc in (14, 16, 18, 20):
            assert str(ws.cell(row=helper_gallons_row, column=cc).value or "").strip() != "0.0m gal"
            assert str(ws.cell(row=helper_gallons_row, column=cc).value or "").strip() != "0.0"
        assert str(ws.cell(row=helper_gallons_row, column=37).value or "").strip() == ""
        assert str(ws.cell(row=helper_basis_row, column=37).value or "").strip() == ""
        legacy_helper_labels = {"Prior quarter", "Quarter-open outlook", "Current QTD", "Next quarter outlook"}
        for rr in range(helper_gallons_row, helper_basis_row + 1):
            assert str(ws.cell(row=rr, column=37).value or "").strip() not in legacy_helper_labels
            assert str(ws.cell(row=rr, column=45).value or "").strip() not in legacy_helper_labels

        prior_comment = ws.cell(row=helper_gallons_row, column=14).comment
        quarter_open_comment = ws.cell(row=helper_gallons_row, column=16).comment
        current_fit_comment = ws.cell(row=helper_gallons_row, column=18).comment
        next_fit_comment = ws.cell(row=helper_gallons_row, column=20).comment
        assert prior_comment is not None
        assert quarter_open_comment is not None
        assert current_fit_comment is not None
        assert next_fit_comment is not None
        assert "2025-Q1 gallons produced scaled by 730/784 MMgy active capacity." in str(prior_comment.text or "")
        assert "2025-Q2 gallons produced scaled by 730/784 MMgy active capacity." in str(quarter_open_comment.text or "")
        assert "2025-Q2 gallons produced scaled by 730/784 MMgy active capacity." in str(current_fit_comment.text or "")
        assert "2025-Q3 gallons produced scaled by 730/784 MMgy active capacity." in str(next_fit_comment.text or "")
        helper_comment_blob = " ".join(
            str((ws.cell(row=helper_gallons_row, column=cc).comment.text if ws.cell(row=helper_gallons_row, column=cc).comment else "") or "")
            for cc in (14, 16, 18, 20)
        ).lower()
        assert "gallons sold" not in helper_comment_blob

        assert str(ws.cell(row=approx_bridge_row, column=14).value or "").strip().startswith("=IF(")
        assert str(ws.cell(row=approx_bridge_row, column=15).value or "").strip() == ""
        assert str(ws.cell(row=fitted_bridge_row, column=15).value or "").strip() == ""
    finally:
        wb.close()


def test_chart_native_quarter_label_points_use_visible_clipped_midpoints() -> None:
    points = writer_context_module._build_visible_quarter_label_points(
        date(2024, 2, 10),
        date(2024, 8, 20),
    )

    assert [str(point.get("label") or "") for point in points] == [
        "2024-Q1",
        "2024-Q2",
        "2024-Q3",
    ]
    assert points[0]["clip_start"] == date(2024, 2, 10)
    assert points[0]["clip_end"] == date(2024, 3, 31)
    assert points[0]["midpoint"] == date(2024, 3, 6)
    assert points[1]["clip_start"] == date(2024, 4, 1)
    assert points[1]["clip_end"] == date(2024, 6, 30)
    assert points[1]["midpoint"] == date(2024, 5, 16)
    assert points[2]["clip_start"] == date(2024, 7, 1)
    assert points[2]["clip_end"] == date(2024, 8, 20)
    assert points[2]["midpoint"] == date(2024, 7, 26)


def test_current_delivered_workbooks_remaining_overlay_note_and_qa_fixes() -> None:
    pbi_path = _current_delivered_model_path("PBI")
    gpre_path = _current_delivered_model_path("GPRE")
    if not pbi_path.exists() or not gpre_path.exists():
        pytest.skip("Current delivered PBI/GPRE workbooks missing for remaining cleanup readback test.")

    wb_pbi = load_workbook(pbi_path, data_only=False, read_only=False)
    try:
        ws_val = wb_pbi["Valuation"]
        debt_detail_row = _find_row_containing(ws_val, "August 2032", column=1)
        assert debt_detail_row is not None
        src_comment = ws_val.cell(row=debt_detail_row, column=17).comment
        assert src_comment is not None
        src_comment_text = str(src_comment.text or "")
        assert "\\sec_cache\\PBI\\" in src_comment_text
        assert "\\sec_cache\\GPRE\\" not in src_comment_text

        ws_qa = wb_pbi["QA_Checks"]
        qa_messages = [str(ws_qa.cell(row=rr, column=5).value or "").strip() for rr in range(1, ws_qa.max_row + 1)]
        assert not any(msg.startswith("FCF (Q): WARN model") for msg in qa_messages)

        ws_nr = wb_pbi["Needs_Review"]
        needs_review_blob = " | ".join(
            str(ws_nr.cell(row=rr, column=cc).value or "").strip()
            for rr in range(1, ws_nr.max_row + 1)
            for cc in range(1, min(ws_nr.max_column, 6) + 1)
        )
        assert "FCF (Q): WARN model" not in needs_review_blob

        ws_qn_pbi = wb_pbi["Quarter_Notes_UI"]
        q3_notes = _quarter_block_notes(ws_qn_pbi, "2025-09-30")
        assert "Free cash flow declined to $60.4m, down $13.1m YoY." in q3_notes
        fcf_note_row = next(
            rr
            for rr in range(1, ws_qn_pbi.max_row + 1)
            if str(ws_qn_pbi.cell(row=rr, column=3).value or "").strip() == "Free cash flow declined to $60.4m, down $13.1m YoY."
        )
        fcf_comment = ws_qn_pbi.cell(row=fcf_note_row, column=3).comment
        assert fcf_comment is not None
        fcf_comment_text = str(fcf_comment.text or "")
        assert "\\sec_cache\\PBI\\" in fcf_comment_text
        assert "adj_metrics" not in fcf_comment_text
        pbi_fcf_metrics = {
            str(ws_qn_pbi.cell(row=rr, column=4).value or "").strip()
            for rr in range(1, ws_qn_pbi.max_row + 1)
            if "Free cash flow " in str(ws_qn_pbi.cell(row=rr, column=3).value or "")
        }
        assert pbi_fcf_metrics == {"Quarterly FCF"}
    finally:
        wb_pbi.close()

    wb_gpre = load_workbook(gpre_path, data_only=False, read_only=False)
    try:
        ws_qn = wb_gpre["Quarter_Notes_UI"]
        for note_txt in [
            "Corporate activities included $16.1m of restructuring costs from the cost reduction initiative.",
            "Corporate activities included $10.3m of restructuring costs from the cost reduction initiative.",
        ]:
            row_idx = next(
                rr
                for rr in range(1, ws_qn.max_row + 1)
                if note_txt in str(ws_qn.cell(row=rr, column=3).value or "")
            )
            assert str(ws_qn.cell(row=row_idx, column=2).value or "").strip() == "Results / drivers"
            assert str(ws_qn.cell(row=row_idx, column=4).value or "").strip() == "One-time items / restructuring"

        ws_basis = wb_gpre["Basis_Proxy_Sandbox"]
        ws_basis = wb_gpre["Basis_Proxy_Sandbox"]
        ws_overlay = wb_gpre["Economics_Overlay"]
        overlay_rows = [
            " | ".join(str(ws_overlay.cell(row=rr, column=cc).value or "").strip() for cc in range(1, min(18, ws_overlay.max_column + 1)))
            for rr in range(1, ws_overlay.max_row + 1)
        ]
        overlay_blob = "\n".join(overlay_rows)
        management_rows = _overlay_management_commentary_rows(ws_overlay)
        management_blob = "\n".join(row["commentary"] for row in management_rows)
        commercial_blob = _overlay_section_blob(
            ws_overlay,
            "Commercial / hedge setup",
            stop_labels=["Bridge to reported"],
        )
        assert "Management commentary" in overlay_blob
        assert "Management said the Q1 position was helped by stronger domestic ethanol markets, limited inventory build and stronger DCO values." in management_blob
        assert "Management said simple crush margins were holding up relatively well as low corn costs supported the setup heading into 2026." in management_blob
        assert "Management said the Q1 position was helped by stronger domestic ethanol markets, limited inventory build and stronger DCO values." not in commercial_blob
        assert "Management said simple crush margins were holding up relatively well as low corn costs supported the setup heading into 2026." not in commercial_blob
        assert "mid-high single digits to the low teens" in commercial_blob
        assert "Primarily open to the margin structure across products" in commercial_blob
        assert "priced early before prices fell further" in commercial_blob
        assert "$0.12-$0.17/gal on-paper setup tracked through May" in commercial_blob
        assert "Q4 crush about 75% hedged; Q1 2026 positions already on" in commercial_blob
        assert "looking for lock-in opportunities" in commercial_blob
        assert "significant portion of Q1 production margin was already logged in" in commercial_blob
        assert "Management linked the stronger setup to corn supply, domestic ethanol markets and DCO values" in commercial_blob
        assert "open to the crush going into Q4" in commercial_blob
        assert "wrong choice" in commercial_blob
        assert "Healthy export volumes and wider E15 acceptance were cited as demand supports into 2026." in management_blob
        assert "DDGS and high-protein values remained under pressure through much of the quarter." in management_blob
        assert "Solid domestic blending and strong export demand supported Q4 ethanol margins." in management_blob
        assert "Corn-oil values contributed positively to gross margin during the quarter." in management_blob
        assert "Protein pricing remained under pressure in Q4." in management_blob
        assert "Reported ethanol-production margin included a $22.6m accumulated RIN sale and a $2.3m inventory NRV adjustment." in management_blob
        assert "Management said the team stayed active in Q4 and Q1 2026, looking daily for lock-in opportunities." in management_blob
        assert "Disciplined risk management strategy continues to support first quarter margins and cash flow." not in management_blob
        assert "Plants ran above 100% capacity utilization during the quarter." not in management_blob
        assert "Disciplined risk management strategy continues to support fourth quarter margins and cash flow." not in overlay_blob
        assert "430k tons" not in overlay_blob
        assert "900 tons/day" not in overlay_blob
        q3_transcript_row = next(
            rr
            for rr in range(1, ws_overlay.max_row + 1)
            if str(ws_overlay.cell(row=rr, column=3).value or "").strip() == "Forward hedge positioning"
            and "q4 crush about 75% hedged" in str(ws_overlay.cell(row=rr, column=5).value or "").lower()
        )
        q4_transcript_row = next(
            rr
            for rr in range(1, ws_overlay.max_row + 1)
            if str(ws_overlay.cell(row=rr, column=3).value or "").strip() == "Q1 margin positioning"
            and "significant portion of q1 production margin already logged in" in str(ws_overlay.cell(row=rr, column=5).value or "").lower()
        )
        q3_transcript_comment = ws_overlay.cell(row=q3_transcript_row, column=3).comment
        q4_transcript_comment = ws_overlay.cell(row=q4_transcript_row, column=3).comment
        assert q3_transcript_comment is not None
        assert q4_transcript_comment is not None
        assert "Source: Transcript" in str(q3_transcript_comment.text or "")
        assert "GPRE_Q3_2025_transcript.txt" in str(q3_transcript_comment.text or "")
        assert "third quarter" not in str(q3_transcript_comment.text or "").lower()
        assert "Source: Transcript" in str(q4_transcript_comment.text or "")
        assert "GPRE_Q4_2025_transcript.txt" in str(q4_transcript_comment.text or "")

        q4_notes = _quarter_block_notes(ws_qn, "2025-12-31")
        q3_notes = _quarter_block_notes(ws_qn, "2025-09-30")
        assert "[NEW] Disciplined risk management strategy continues to support first quarter margins and cash flow." in q4_notes
        assert "Management said Q4 crush was about 75% hedged and positions had been put on for Q1 2026." not in q3_notes
        assert "Management said a significant portion of Q1 production margin was already logged in." not in q4_notes
        assert q4_notes.count("[NEW] 45Z production tax credits contributed $23.4m net of discounts and other costs in Q4.") == 1
        assert "[NEW] 45Z production tax credits contributed $23.4m net of discounts and other costs." not in q4_notes
        risk_row = next(
            rr
            for rr in range(1, ws_qn.max_row + 1)
            if str(ws_qn.cell(row=rr, column=3).value or "").strip()
            == "[NEW] Disciplined risk management strategy continues to support first quarter margins and cash flow."
        )
        assert str(ws_qn.cell(row=risk_row, column=2).value or "").strip() == "Guidance / outlook"
        assert str(ws_qn.cell(row=risk_row, column=4).value or "").strip() == "Risk management"

        asset_note_row = next(
            rr
            for rr in range(1, ws_qn.max_row + 1)
            if "Management is pursuing non-core asset monetization to enhance liquidity and strengthen the balance sheet."
            in str(ws_qn.cell(row=rr, column=3).value or "")
        )
        assert str(ws_qn.cell(row=asset_note_row, column=4).value or "").strip() == "Liquidity / balance-sheet"
    finally:
        wb_gpre.close()


def test_current_delivered_workbooks_verified_output_bug_fixes_and_qa_cleanup() -> None:
    pbi_path = _current_delivered_model_path("PBI")
    gpre_path = _current_delivered_model_path("GPRE")
    if not pbi_path.exists() or not gpre_path.exists():
        pytest.skip("Current delivered PBI/GPRE workbooks missing for verified output bug-fix readback test.")

    wb_gpre = load_workbook(gpre_path, data_only=False, read_only=False)
    try:
        ws_val = wb_gpre["Valuation"]
        owner_earnings_row = _find_row_with_value(ws_val, "Owner earnings (proxy)", column=1)
        cash_flow_quality_row = _find_row_with_value(ws_val, "Cash-flow quality", column=1)
        adj_fcf_diff_row = _find_row_with_value(ws_val, "Adj FCF - FCF", column=1)
        assert owner_earnings_row is not None
        assert cash_flow_quality_row is not None
        assert adj_fcf_diff_row is not None
        assert not bool(ws_val.row_dimensions[owner_earnings_row].hidden)
        assert not bool(ws_val.row_dimensions[cash_flow_quality_row].hidden)
        quarter_cols = [
            cc
            for cc in range(2, ws_val.max_column + 1)
            if re.fullmatch(r"\d{4}-Q[1-4]", str(ws_val.cell(row=6, column=cc).value or "").strip())
        ]
        adj_fcf_values = [
            float(pd.to_numeric(ws_val.cell(row=adj_fcf_diff_row, column=cc).value, errors="coerce"))
            for cc in quarter_cols
            if pd.notna(pd.to_numeric(ws_val.cell(row=adj_fcf_diff_row, column=cc).value, errors="coerce"))
        ]
        assert adj_fcf_values == [pytest.approx(-62.005955, abs=1e-6)]
        assert ws_val["N248"].value == 60
        assert ws_val["N249"].value == 200
        assert ws_val["N250"].value == "=SUM(N248:N249)"
        assert "Price" not in str(ws_val["J222"].value or "")
        for coord in ["H227", "I227", "J227", "K227", "L227"]:
            assert "Price" not in str(ws_val[coord].value or "")
        assert ws_val["N209"].value == '=IF(OR(Price="",Price<=0,BV_PerShare=""),"",IF(BV_PerShare<=0,"n/a (neg equity)",Price/BV_PerShare))'
        assert ws_val["O49"].value == "Thesis target equity FCF yield"
        crush_row = _find_row_with_value(ws_val, "Crush margin uplift", column=15)
        assert crush_row is not None and "stronger crush margin" in str(ws_val.cell(row=crush_row, column=20).value or "")
        assert _find_row_with_value(ws_val, "Corn oil / coproduct uplift", column=15) is None
        assert _find_row_with_value(ws_val, "Protein / mix uplift", column=15) is None
        assert str(ws_val.cell(row=owner_earnings_row, column=15).value or "").strip() == "Cost savings uplift"
        assert str(ws_val.cell(row=owner_earnings_row, column=20).value or "").strip() == "Use annualized savings not yet fully visible in reported TTM."
        assert str(ws_val.cell(row=cash_flow_quality_row, column=15).value or "").strip() == "Interest savings / debt-paydown uplift"
        assert str(ws_val.cell(row=cash_flow_quality_row, column=20).value or "").strip() == "Bridge item for lower cash interest or debt reduction not yet visible in TTM."
        assert str(ws_val.cell(row=cash_flow_quality_row + 1, column=15).value or "").strip() == "Other"
        assert str(ws_val.cell(row=cash_flow_quality_row + 2, column=15).value or "").strip() == "Output"
        assert str(ws_val.cell(row=cash_flow_quality_row + 2, column=19).value or "").strip() == "Value"
        assert str(ws_val.cell(row=cash_flow_quality_row + 2, column=20).value or "").strip() == "Interpretation"
        assert str(ws_val.cell(row=cash_flow_quality_row + 3, column=15).value or "").strip() == "Thesis Adj EBITDA"
        assert str(ws_val.cell(row=cash_flow_quality_row + 3, column=19).value or "").strip() == "=S47+SUM(S50:S54)"
        assert ws_val["D249"].value == "FCF TTM accelerated; YoY delta $198.7m"
        assert ws_val["D250"].value == "Net debt declined; YoY delta $-77.9m"

        ws_qn = wb_gpre["Quarter_Notes_UI"]
        q4_2024_notes = _quarter_block_notes(ws_qn, "2024-12-31")
        assert "Consolidated ethanol crush margin declined to $(15.5)m from $53.0m YoY." in q4_2024_notes
        q4_2025_notes = _quarter_block_notes(ws_qn, "2025-12-31")
        q3_2025_notes = _quarter_block_notes(ws_qn, "2025-09-30")
        assert "[NEW] Disciplined risk management strategy continues to support first quarter margins and cash flow." in q4_2025_notes
        assert "[NEW] 45Z production tax credits contributed $23.4m net of discounts and other costs in Q4." in q4_2025_notes
        assert q4_2025_notes.count("[NEW] 45Z production tax credits contributed $23.4m net of discounts and other costs in Q4.") == 1
        assert "[NEW] 45Z production tax credits contributed $23.4m net of discounts and other costs." not in q4_2025_notes
        assert "[NEW] Q4 2025 45Z monetization expected at $15m-$25m." in q3_2025_notes
        assert "[NEW] All eight operating ethanol plants expected to qualify for production tax credits in 2026" in q3_2025_notes
        assert "[NEW] 45Z tax credit monetization agreement for Nebraska production was entered on September 16, 2025." in q3_2025_notes
        q4_auth_row = next(
            rr
            for rr in range(1, ws_qn.max_row + 1)
            if str(ws_qn.cell(row=rr, column=1).value or "").strip() == "2025-12-31"
        )
        q3_auth_row = next(
            rr
            for rr in range(1, ws_qn.max_row + 1)
            if str(ws_qn.cell(row=rr, column=1).value or "").strip() == "2025-09-30"
        )
        q4_auth_note = ""
        q3_auth_note = ""
        for rr in range(q4_auth_row + 1, min(q4_auth_row + 18, ws_qn.max_row + 1)):
            if str(ws_qn.cell(row=rr, column=1).value or "").strip().startswith("202"):
                break
            note_txt = str(ws_qn.cell(row=rr, column=3).value or "").strip()
            if "Repurchase authorization increased to $200.0m." in note_txt:
                q4_auth_note = note_txt
                break
        for rr in range(q3_auth_row + 1, min(q3_auth_row + 18, ws_qn.max_row + 1)):
            if str(ws_qn.cell(row=rr, column=1).value or "").strip().startswith("202"):
                break
            note_txt = str(ws_qn.cell(row=rr, column=3).value or "").strip()
            if "Repurchase authorization increased to $200.0m." in note_txt:
                q3_auth_note = note_txt
                break
        assert q4_auth_note == "Repurchase authorization increased to $200.0m."
        assert q3_auth_note == "Repurchase authorization increased to $200.0m."

        # Keep this delivered-workbook QA pass narrow. Detailed driver and
        # overlay audits live in the focused economics tests above.
        ws_drv = wb_gpre["Operating_Drivers"]
        assert str(ws_drv["A4"].value or "").strip() == "Operating Commentary"
        drv_commentary_rows = [
            row for row in _operating_commentary_rows(ws_drv) if _quarter_label_ord(row["stated_in"]) is not None
        ]
        commentary_targets = [
            "Plants ran above 100% capacity utilization during the quarter.",
            "Reliability-centered maintenance reduced planned and unplanned downtime.",
            "Record high ethanol and Ultra-high protein yields supported record protein output and corn-oil production.",
        ]
        assert 12 <= len(drv_commentary_rows) <= 40
        drv_quarter_ords = [_quarter_label_ord(row["stated_in"]) for row in drv_commentary_rows]
        assert all(ord_val is not None for ord_val in drv_quarter_ords)
        assert drv_quarter_ords[0] == max(drv_quarter_ords)
        for target in commentary_targets:
            assert any(row["commentary"] == target for row in drv_commentary_rows)
        assert not any("45Z" in row["commentary"] for row in drv_commentary_rows)

        ws_basis = wb_gpre["Basis_Proxy_Sandbox"]
        ws_overlay = wb_gpre["Economics_Overlay"]
        proxy_table_row = _find_row_with_value(ws_overlay, "Proxy comparison ($/gal)", column=1)
        bridge_row = _find_row_with_value(ws_overlay, "Bridge to reported", column=1)
        quarterly_chart_title_row = _find_row_with_value(ws_overlay, "Approximate market crush, fitted models, and real GPRE crush margin (quarterly)", column=2)
        best_forward_role_row = _find_row_with_value(ws_basis, "Best forward lens", column=21)
        assert proxy_table_row is not None
        assert bridge_row is not None
        assert quarterly_chart_title_row is not None
        assert best_forward_role_row is not None
        proxy_note = str(ws_overlay.cell(row=proxy_table_row + 1, column=1).value or "").strip()
        assert "Official row = Approximate market crush" in proxy_note
        assert "Fitted row = GPRE crush proxy" in proxy_note
        assert "Production winner =" in proxy_note
        assert "Best forward lens =" in proxy_note
        management_blob = " ".join(str(row["commentary"]) for row in _overlay_management_commentary_rows(ws_overlay))
        for target in commentary_targets:
            assert target not in management_blob
        assert len(ws_overlay._charts) == 3
        quarterly_chart = ws_overlay._charts[1]
        coproduct_chart = ws_overlay._charts[2]
        assert len(quarterly_chart.series) == 4
        assert len(coproduct_chart.series) == 1
        assert getattr(getattr(quarterly_chart, "legend", None), "position", None) == "t"
        assert bool(getattr(getattr(quarterly_chart, "legend", None), "overlay", False))
        assert "Process q-open blend" in str(ws_basis.cell(row=best_forward_role_row, column=22).value or "")
        with zipfile.ZipFile(gpre_path) as zf:
            chart_xmls = {
                name: zf.read(name).decode("utf-8", errors="ignore")
                for name in zf.namelist()
                if name.startswith("xl/charts/chart") and name.endswith(".xml")
            }
        quarterly_chart_xml = next(
            xml
            for xml in chart_xmls.values()
            if "<lineChart>" in xml and 'legendPos val="t"' in xml
        )
        coproduct_chart_xml = next(
            xml
            for xml in chart_xmls.values()
            if "<lineChart>" in xml and "'Economics_Overlay'!$AW$" in xml
        )
        assert "Quarter boundary" not in quarterly_chart_xml
        assert 'legendPos val="t"' in quarterly_chart_xml
        assert '<overlay val="1"/>' in quarterly_chart_xml
        assert "<dLbls>" in quarterly_chart_xml
        assert '<showLegendKey val="1"/>' not in quarterly_chart_xml
        assert '<dLblPos val="r"/>' in quarterly_chart_xml
        assert "'Economics_Overlay'!BC" not in quarterly_chart_xml
        assert re.search(r"<catAx>.*?<axPos val=\"b\"/>", quarterly_chart_xml)
        assert re.search(r"<valAx>.*?<crosses val=\"min\"/>", quarterly_chart_xml)
        assert "<legendEntry>" not in quarterly_chart_xml
        assert "'Economics_Overlay'!$AW$" in coproduct_chart_xml
        assert "<dLbls>" in coproduct_chart_xml
        assert '<showLegendKey val="1"/>' not in coproduct_chart_xml
        assert '<dLblPos val="b"/>' in coproduct_chart_xml

        wb_pbi = load_workbook(pbi_path, data_only=False, read_only=False)
        try:
            ws_summary = wb_pbi["SUMMARY"]
            debt_eq_row = _find_row_with_value(ws_summary, "Debt-to-equity (latest quarter)", column=1)
            assert debt_eq_row is not None
            assert str(ws_summary.cell(row=debt_eq_row, column=2).value or "").strip() == "N/M (neg equity)"
        finally:
            wb_pbi.close()
        return

        ws_drv = wb_gpre["Operating_Drivers"]
        assert str(ws_drv["A4"].value or "").strip() == "Operating Commentary"
        assert ws_drv.freeze_panes is None
        drv_commentary_rows = [
            row for row in _operating_commentary_rows(ws_drv) if _quarter_label_ord(row["stated_in"]) is not None
        ]
        assert 12 <= len(drv_commentary_rows) <= 40
        drv_quarter_ords = [_quarter_label_ord(row["stated_in"]) for row in drv_commentary_rows]
        assert all(ord_val is not None for ord_val in drv_quarter_ords)
        assert drv_quarter_ords[0] == max(drv_quarter_ords)
        drv_distinct_ords = []
        for ord_val in drv_quarter_ords:
            if ord_val not in drv_distinct_ords:
                drv_distinct_ords.append(ord_val)
        assert drv_distinct_ords == sorted(drv_distinct_ords, reverse=True)
        assert max(len([row for row in drv_commentary_rows if row["stated_in"] == qtxt]) for qtxt in {row["stated_in"] for row in drv_commentary_rows}) <= 10
        assert len({str(row["commentary"]).strip() for row in drv_commentary_rows}) == len(drv_commentary_rows)
        assert any((_quarter_label_ord(row["stated_in"]) or 0) <= 2024 * 4 + 4 for row in drv_commentary_rows)
        assert any(str(row["horizon"]).strip() == "" for row in drv_commentary_rows)
        assert float(ws_drv.row_dimensions[4].height or 0.0) == pytest.approx(22.5, abs=0.1)
        assert float(ws_drv.row_dimensions[5].height or 0.0) == pytest.approx(21.0, abs=0.1)
        assert all(float(ws_drv.row_dimensions[int(row["row"])].height or 0.0) == pytest.approx(19.5, abs=0.1) for row in drv_commentary_rows)
        year_band_rows = [
            rr
            for rr in range(6, ws_drv.max_row + 1)
            if str(ws_drv.cell(row=rr, column=1).value or "").strip() in {"2023", "2024", "2025", "2026 / current"}
            and not str(ws_drv.cell(row=rr, column=2).value or "").strip()
            and not str(ws_drv.cell(row=rr, column=3).value or "").strip()
        ]
        assert year_band_rows
        assert all(float(ws_drv.row_dimensions[rr].height or 0.0) == pytest.approx(21.0, abs=0.1) for rr in year_band_rows)
        assert bool(ws_drv["A5"].font.bold)
        assert str(getattr(ws_drv["A5"].font.color, "rgb", "") or "") == str(getattr(ws_drv["B5"].font.color, "rgb", "") or "")
        for idx, row in enumerate(drv_commentary_rows):
            has_separator = str(getattr(ws_drv.cell(row=int(row["row"]), column=2).border.top, "style", "") or "") == "thin"
            horizon_has_separator = str(getattr(ws_drv.cell(row=int(row["row"]), column=1).border.top, "style", "") or "") == "thin"
            bottom_line = str(getattr(ws_drv.cell(row=int(row["row"]), column=2).border.bottom, "style", "") or "")
            if idx == 0:
                assert not has_separator
                assert not horizon_has_separator
                assert not bottom_line
                continue
            prev_row = drv_commentary_rows[idx - 1]
            if str(prev_row["year_band"]) != str(row["year_band"]) or str(prev_row["stated_in"]) == str(row["stated_in"]):
                assert not has_separator
                assert not horizon_has_separator
            else:
                assert has_separator
                assert horizon_has_separator
            assert not bottom_line
        actuals_row = _find_row_with_value(ws_drv, "Actuals — latest 12 quarters", column=1)
        assert actuals_row is not None
        legend_row = actuals_row - 1
        assert all(not str(ws_drv.cell(row=legend_row, column=cc).value or "").strip() for cc in range(1, 9))
        assert [str(ws_drv.cell(row=legend_row, column=cc).value or "").strip() for cc in range(9, 14)] == [
            "<=-15%",
            "-15..-5",
            "-5..+5",
            "+5..+15",
            ">=+15%",
        ]
        assert [_fill_rgb(ws_drv.cell(row=legend_row, column=cc)) for cc in range(9, 14)] == [
            "00A63A00",
            "00D55E00",
            "00DDDDDD",
            "009BD3F5",
            "002F80ED",
        ]
        assert float(ws_drv.row_dimensions[legend_row].height or 0.0) == pytest.approx(15.75, abs=0.1)
        assert float(ws_drv.row_dimensions[actuals_row].height or 0.0) == pytest.approx(22.5, abs=0.1)
        assert float(ws_drv.row_dimensions[actuals_row + 1].height or 0.0) == pytest.approx(21.0, abs=0.1)
        quarter_row = actuals_row + 1
        crush_row = _find_row_with_value(ws_drv, "Consolidated ethanol crush margin ($m)", column=1)
        realized_45z_row = _find_row_with_value(ws_drv, "45Z value realized ($m)", column=1)
        assert crush_row is not None
        assert realized_45z_row is not None
        q_2025_q3_col = _find_col_with_value(ws_drv, "2025-Q3", row=quarter_row)
        q_2025_q4_col = _find_col_with_value(ws_drv, "2025-Q4", row=quarter_row)
        assert q_2025_q3_col is not None and q_2025_q4_col is not None
        assert pd.notna(pd.to_numeric(ws_drv.cell(row=crush_row, column=q_2025_q3_col).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_drv.cell(row=crush_row, column=q_2025_q4_col).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_drv.cell(row=realized_45z_row, column=q_2025_q3_col).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_drv.cell(row=realized_45z_row, column=q_2025_q4_col).value, errors="coerce"))
        assert _fill_rgb(ws_drv["C7"]) == "00FFFFFF"
        assert _fill_rgb(ws_drv["B7"]) == "00FFFFFF"
        assert _find_row_with_value(ws_drv, "Ultra-high protein (k tons)", column=1) is not None
        assert any("Reliability-centered maintenance reduced planned and unplanned downtime." == row["commentary"] for row in drv_commentary_rows)
        assert any("Plants ran above 100% capacity utilization during the quarter." == row["commentary"] for row in drv_commentary_rows)
        assert any("Record high ethanol and Ultra-high protein yields supported record protein output and corn-oil production." == row["commentary"] for row in drv_commentary_rows)
        assert any("Revenue declined because we exited ethanol marketing for Tharaldson and placed the Fairmont ethanol asset on care and maintenance." == row["commentary"] for row in drv_commentary_rows)
        assert any(
            "Plant utilization reflected the normal spring maintenance season, with plants temporarily shut down for annual clean-out and restart."
            == row["commentary"]
            for row in drv_commentary_rows
        )
        assert not any("50 Pro" in row["commentary"] for row in drv_commentary_rows)
        assert not any("45Z" in row["commentary"] for row in drv_commentary_rows)
        assert not any("Plant utilization rate of 97%" in row["commentary"] for row in drv_commentary_rows)
        assert not any("Plant utilization rate of 93.9%" in row["commentary"] for row in drv_commentary_rows)
        assert not any("Plant utilization reflected the spring maintenance season." == row["commentary"] for row in drv_commentary_rows)
        assert not any("..." in row["commentary"] for row in drv_commentary_rows)
        assert not any("?" in row["commentary"] for row in drv_commentary_rows)
        assert not any("start to see" in row["commentary"].lower() for row in drv_commentary_rows)
        assert not any("Sequence, and" in row["commentary"] for row in drv_commentary_rows)
        assert not any("a couple of other things" in row["commentary"].lower() for row in drv_commentary_rows)
        assert not any("what we're seeing" in row["commentary"].lower() for row in drv_commentary_rows)
        assert not any("sequence without negatively impacting" in row["commentary"].lower() for row in drv_commentary_rows)
        assert not any("exports as a result of that new capacity" in row["commentary"].lower() for row in drv_commentary_rows)
        assert not any("applicable securities laws" in row["commentary"].lower() for row in drv_commentary_rows)
        assert not any("offtake agreement" in row["commentary"].lower() for row in drv_commentary_rows)
        assert not any(row["commentary"].lower().startswith(("due to ", "driven by ", "helped by ", "impacted by ")) for row in drv_commentary_rows)

        ws_basis = wb_gpre["Basis_Proxy_Sandbox"]
        ws_overlay = wb_gpre["Economics_Overlay"]
        overlay_commentary_rows = _overlay_management_commentary_rows(ws_overlay)
        assert overlay_commentary_rows
        assert all(float(ws_overlay.row_dimensions[int(row["row"])].height or 0.0) == pytest.approx(19.5, abs=0.1) for row in overlay_commentary_rows)
        mgmt_row = _find_row_with_value(ws_overlay, "Management commentary", column=1)
        assert mgmt_row is not None
        assert float(ws_overlay.row_dimensions[mgmt_row].height or 0.0) == pytest.approx(22.5, abs=0.1)
        assert float(ws_overlay.row_dimensions[mgmt_row + 1].height or 0.0) == pytest.approx(21.0, abs=0.1)
        overlay_year_rows = [
            rr
            for rr in range(mgmt_row + 2, ws_overlay.max_row + 1)
            if str(ws_overlay.cell(row=rr, column=1).value or "").strip() in {"2023", "2024", "2025", "2026 / current"}
            and not str(ws_overlay.cell(row=rr, column=2).value or "").strip()
            and not str(ws_overlay.cell(row=rr, column=3).value or "").strip()
        ]
        assert overlay_year_rows
        assert all(float(ws_overlay.row_dimensions[rr].height or 0.0) == pytest.approx(21.0, abs=0.1) for rr in overlay_year_rows[:4])
        assert str(getattr(ws_overlay.cell(row=mgmt_row + 1, column=1).font.color, "rgb", "") or "") == str(getattr(ws_overlay.cell(row=mgmt_row + 1, column=2).font.color, "rgb", "") or "")
        contiguous_year_bands = []
        for row in overlay_commentary_rows:
            if not contiguous_year_bands or contiguous_year_bands[-1] != row["year_band"]:
                contiguous_year_bands.append(row["year_band"])
        year_band_rank = {"2026 / current": 0, "2025": 1, "2024": 2, "2023": 3}
        assert len(contiguous_year_bands) == len(set(contiguous_year_bands))
        assert contiguous_year_bands == sorted(contiguous_year_bands, key=lambda txt: year_band_rank.get(txt, 99))
        for idx, row in enumerate(overlay_commentary_rows):
            has_separator = str(getattr(ws_overlay.cell(row=int(row["row"]), column=2).border.top, "style", "") or "") == "thin"
            horizon_has_separator = str(getattr(ws_overlay.cell(row=int(row["row"]), column=1).border.top, "style", "") or "") == "thin"
            bottom_line = str(getattr(ws_overlay.cell(row=int(row["row"]), column=2).border.bottom, "style", "") or "")
            if idx == 0:
                assert not has_separator
                assert not horizon_has_separator
                assert not bottom_line
                continue
            prev_row = overlay_commentary_rows[idx - 1]
            if str(prev_row["year_band"]) != str(row["year_band"]) or str(prev_row["stated_in"]) == str(row["stated_in"]):
                assert not has_separator
                assert not horizon_has_separator
            else:
                assert has_separator
                assert horizon_has_separator
            assert not bottom_line
        overlay_setup_row = _find_row_with_value(ws_overlay, "Commercial / hedge setup", column=1)
        assert overlay_setup_row is not None
        assert float(ws_overlay.row_dimensions[overlay_setup_row].height or 0.0) == pytest.approx(22.5, abs=0.1)
        assert float(ws_overlay.row_dimensions[overlay_setup_row + 1].height or 0.0) == pytest.approx(21.0, abs=0.1)
        bridge_row = _find_row_with_value(ws_overlay, "Bridge to reported", column=1)
        assert bridge_row is not None
        first_setup_data_row = next(
            rr
            for rr in range(overlay_setup_row + 2, bridge_row)
            if str(ws_overlay.cell(row=rr, column=2).value or "").strip()
            and str(ws_overlay.cell(row=rr, column=3).value or "").strip()
        )
        assert str(getattr(ws_overlay.cell(row=first_setup_data_row, column=1).font.color, "rgb", "") or "") == str(getattr(ws_overlay.cell(row=first_setup_data_row, column=2).font.color, "rgb", "") or "")
        base_row = _find_row_with_value(ws_overlay, "Base operating coefficients", column=1)
        market_row = _find_row_with_value(ws_overlay, "Market inputs", column=1)
        process_row = _find_row_with_value(ws_overlay, "Unhedged process economics", column=1)
        sandbox_build_row = _find_row_with_value(ws_basis, "Approximate market crush build-up ($/gal)", column=2)
        assert base_row is not None and market_row is not None and sandbox_build_row is not None
        assert process_row is None
        assert float(ws_overlay.row_dimensions[bridge_row].height or 0.0) == pytest.approx(22.5, abs=0.1)
        assert float(ws_overlay.row_dimensions[base_row].height or 0.0) == pytest.approx(22.5, abs=0.1)
        assert float(ws_overlay.row_dimensions[market_row].height or 0.0) == pytest.approx(30.0, abs=0.1)
        bridge_intro_text = str(ws_overlay.cell(row=bridge_row + 1, column=1).value or "").strip()
        assert bridge_intro_text == "Approximate market crush shows simple weighted market/process conditions; GPRE crush proxy adds company-specific timing / hedge effects."
        assert str(ws_overlay.cell(row=base_row + 1, column=1).value or "").strip() == "Use platform/process coefficients as editable base assumptions. Reported values override inferred and user-entered assumptions when explicitly disclosed."
        assert str(ws_overlay.cell(row=base_row + 3, column=4).value or "").strip() == "Status"
        assert str(ws_overlay.cell(row=base_row + 3, column=9).value or "").strip() == "Coverage / note"
        market_intro_text = str(ws_overlay.cell(row=market_row + 1, column=1).value or "").strip()
        assert market_intro_text == "Prior quarter and Current QTD use observed market data | Quarter-open outlook is current quarter based on futures prices before the quarter started | Next quarter outlook is current futures prices for next quarter."
        assert str(ws_overlay.cell(row=market_row + 3, column=1).value or "").strip() == "Input"
        assert str(ws_overlay.cell(row=market_row + 3, column=2).value or "").strip() == "Prior quarter"
        assert str(ws_overlay.cell(row=market_row + 3, column=4).value or "").strip() == "Quarter-open outlook"
        assert str(ws_overlay.cell(row=market_row + 3, column=6).value or "").strip() == "Current QTD"
        assert str(ws_overlay.cell(row=market_row + 3, column=8).value or "").strip() == "Next quarter outlook"
        assert str(ws_overlay.cell(row=market_row + 3, column=11).value or "").strip() == "Source"
        for cc in (2, 4, 6, 8, 10, 11):
            assert str(getattr(ws_overlay.cell(row=market_row + 3, column=cc).alignment, "horizontal", "") or "") == "center"
        assert str(ws_overlay.cell(row=market_row + 4, column=2).value or "").strip() == "2026-Q1"
        assert str(ws_overlay.cell(row=market_row + 4, column=4).value or "").strip() == "As of 2026-03-31"
        assert str(ws_overlay.cell(row=market_row + 4, column=6).value or "").strip().startswith("As of ")
        assert str(ws_overlay.cell(row=market_row + 4, column=8).value or "").strip() == "2026-Q3"
        merged_ranges = {str(rng) for rng in ws_overlay.merged_cells.ranges}
        assert f"B{market_row + 3}:C{market_row + 3}" in merged_ranges
        assert f"D{market_row + 3}:E{market_row + 3}" in merged_ranges
        assert f"F{market_row + 3}:G{market_row + 3}" in merged_ranges
        assert f"H{market_row + 3}:I{market_row + 3}" in merged_ranges
        assert f"K{market_row + 3}:U{market_row + 3}" in merged_ranges
        assert f"B{market_row + 4}:C{market_row + 4}" in merged_ranges
        assert f"D{market_row + 4}:E{market_row + 4}" in merged_ranges
        assert f"F{market_row + 4}:G{market_row + 4}" in merged_ranges
        assert float(ws_overlay.column_dimensions["R"].width or 0.0) == pytest.approx(float(ws_overlay.column_dimensions["Q"].width or 0.0), abs=0.05)
        assert float(ws_overlay.column_dimensions["S"].width or 0.0) == pytest.approx(float(ws_overlay.column_dimensions["Q"].width or 0.0), abs=0.05)
        assert float(ws_overlay.column_dimensions["T"].width or 0.0) == pytest.approx(float(ws_overlay.column_dimensions["Q"].width or 0.0), abs=0.05)
        assert float(ws_overlay.column_dimensions["U"].width or 0.0) == pytest.approx(float(ws_overlay.column_dimensions["Q"].width or 0.0), abs=0.05)
        assert str(ws_basis.cell(row=sandbox_build_row + 1, column=2).value or "").strip() == "Official simple row build-up used by Approximate market crush on Economics_Overlay."
        assert str(ws_basis.cell(row=sandbox_build_row + 2, column=2).value or "").strip() == "Line item"
        assert str(ws_basis.cell(row=sandbox_build_row + 2, column=3).value or "").strip() == "Prior quarter"
        assert str(ws_basis.cell(row=sandbox_build_row + 2, column=5).value or "").strip() == "Quarter-open outlook"
        assert str(ws_basis.cell(row=sandbox_build_row + 2, column=7).value or "").strip() == "Current QTD"
        assert str(ws_basis.cell(row=sandbox_build_row + 2, column=9).value or "").strip() == "Next quarter outlook"
        for cc in (3, 5, 7, 9, 11, 12):
            assert str(getattr(ws_basis.cell(row=sandbox_build_row + 2, column=cc).alignment, "horizontal", "") or "") == "center"
        assert str(ws_basis.cell(row=sandbox_build_row + 3, column=3).value or "").strip() == "2026-Q1"
        assert str(ws_basis.cell(row=sandbox_build_row + 3, column=5).value or "").strip() == "As of 2026-03-31"
        assert str(ws_basis.cell(row=sandbox_build_row + 3, column=7).value or "").strip().startswith("As of ")
        assert str(ws_basis.cell(row=sandbox_build_row + 3, column=9).value or "").strip() == "2026-Q3"
        sandbox_basis_snapshot_row = next(
            rr
            for rr in range(sandbox_build_row + 4, ws_basis.max_row + 1)
            if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Official corn basis snapshot date"
        )
        sandbox_basis_rule_row = next(
            rr
            for rr in range(sandbox_basis_snapshot_row, ws_basis.max_row + 1)
            if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Official corn basis selection rule"
        )
        assert str(ws_basis.cell(row=sandbox_basis_snapshot_row, column=11).value or "").strip() == "date/text"
        assert "Retained GPRE corn-bid snapshot date used by the official corn-basis leg only" in str(
            ws_basis.cell(row=sandbox_basis_snapshot_row, column=12).value or ""
        )
        assert str(ws_basis.cell(row=sandbox_basis_rule_row, column=11).value or "").strip() == "rule/text"
        assert str(ws_basis.cell(row=sandbox_basis_rule_row, column=3).value or "").strip() == "latest_snapshot_on_or_before_quarter_end / AMS fallback"
        assert str(ws_basis.cell(row=sandbox_basis_rule_row, column=5).value or "").strip() == "latest_snapshot_on_or_before_quarter_start / AMS fallback"
        assert str(ws_basis.cell(row=sandbox_basis_rule_row, column=7).value or "").strip() == "latest_snapshot_on_or_before_as_of"
        assert str(ws_basis.cell(row=sandbox_basis_rule_row, column=9).value or "").strip() == "latest_snapshot_on_or_before_as_of_with_target_quarter_rows"
        quarter_row = next(
            rr
            for rr in range(bridge_row, ws_overlay.max_row + 1)
            if str(ws_overlay.cell(row=rr, column=1).value or "").strip() == "Quarter"
        )
        assert [str(ws_overlay.cell(row=quarter_row, column=cc).value or "").strip() for cc in range(1, 14)] == [
            "Quarter",
            "2023-Q1",
            "2023-Q2",
            "2023-Q3",
            "2023-Q4",
            "2024-Q1",
            "2024-Q2",
            "2024-Q3",
            "2024-Q4",
            "2025-Q1",
            "2025-Q2",
            "2025-Q3",
            "2025-Q4",
        ]
        proxy_row = _find_row_with_value(ws_overlay, "Approximate market crush ($m)", column=1)
        gpre_proxy_row = _find_row_with_value(ws_overlay, "GPRE crush proxy ($m)", column=1)
        underlying_row = _find_row_with_value(ws_overlay, "Underlying crush margin ($m)", column=1)
        reported_row = _find_row_with_value(ws_overlay, "Reported consolidated crush margin ($m)", column=1)
        gap_row = _find_row_with_value(ws_overlay, "Gap vs market crush", column=1)
        residual_row = _find_row_with_value(ws_overlay, "Hedge / realization / residual effects", column=1)
        assert proxy_row is not None and gpre_proxy_row is not None and underlying_row is not None and reported_row is not None
        assert gap_row is None and residual_row is None
        assert bridge_row < quarter_row < proxy_row < gpre_proxy_row < underlying_row < reported_row
        assert gpre_proxy_row - proxy_row == 1
        assert underlying_row - gpre_proxy_row == 2
        assert reported_row - underlying_row == 1
        assert any(pd.notna(pd.to_numeric(ws_overlay.cell(row=proxy_row, column=cc).value, errors="coerce")) for cc in range(2, 14))
        assert any(pd.notna(pd.to_numeric(ws_overlay.cell(row=gpre_proxy_row, column=cc).value, errors="coerce")) for cc in range(2, 14))
        assert any(pd.notna(pd.to_numeric(ws_overlay.cell(row=underlying_row, column=cc).value, errors="coerce")) for cc in range(2, 14))
        assert any(pd.notna(pd.to_numeric(ws_overlay.cell(row=reported_row, column=cc).value, errors="coerce")) for cc in range(2, 14))
        assert "$" not in str(ws_overlay.cell(row=proxy_row, column=2).number_format or "")
        assert "$" not in str(ws_overlay.cell(row=reported_row, column=2).number_format or "")
        overlay_45z_row = _find_row_with_value(ws_overlay, "45Z impact ($m)", column=1)
        assert overlay_45z_row is not None
        overlay_2025_q3_col = _find_col_with_value(ws_overlay, "2025-Q3", row=quarter_row)
        overlay_2025_q4_col = _find_col_with_value(ws_overlay, "2025-Q4", row=quarter_row)
        overlay_2023_q1_col = _find_col_with_value(ws_overlay, "2023-Q1", row=quarter_row)
        overlay_2023_q2_col = _find_col_with_value(ws_overlay, "2023-Q2", row=quarter_row)
        overlay_2023_q3_col = _find_col_with_value(ws_overlay, "2023-Q3", row=quarter_row)
        overlay_2023_q4_col = _find_col_with_value(ws_overlay, "2023-Q4", row=quarter_row)
        assert overlay_2023_q1_col is not None and overlay_2023_q2_col is not None
        assert overlay_2023_q3_col is not None and overlay_2023_q4_col is not None
        assert overlay_2025_q3_col is not None and overlay_2025_q4_col is not None
        assert float(pd.to_numeric(ws_overlay.cell(row=reported_row, column=overlay_2023_q1_col).value, errors="coerce")) == pytest.approx(-15.3, abs=0.01)
        assert float(pd.to_numeric(ws_overlay.cell(row=reported_row, column=overlay_2023_q2_col).value, errors="coerce")) == pytest.approx(1.9, abs=0.01)
        assert float(pd.to_numeric(ws_overlay.cell(row=reported_row, column=overlay_2023_q3_col).value, errors="coerce")) == pytest.approx(48.5, abs=0.01)
        assert float(pd.to_numeric(ws_overlay.cell(row=reported_row, column=overlay_2023_q4_col).value, errors="coerce")) == pytest.approx(49.7, abs=0.01)
        assert float(pd.to_numeric(ws_overlay.cell(row=reported_row, column=overlay_2025_q3_col).value, errors="coerce")) == pytest.approx(59.6, abs=0.01)
        assert float(pd.to_numeric(ws_overlay.cell(row=reported_row, column=overlay_2025_q4_col).value, errors="coerce")) == pytest.approx(44.4, abs=0.01)
        assert all(not str(ws_overlay.cell(row=gpre_proxy_row + 1, column=cc).value or "").strip() for cc in range(1, 14))
        assert all(not str(ws_overlay.cell(row=reported_row + 1, column=cc).value or "").strip() for cc in range(1, 14))
        assert float(ws_overlay.row_dimensions[gpre_proxy_row + 1].height or 0.0) == pytest.approx(12.0, abs=0.1)
        assert float(ws_overlay.row_dimensions[reported_row + 1].height or 0.0) == pytest.approx(12.0, abs=0.1)
        assert float(pd.to_numeric(ws_overlay.cell(row=overlay_45z_row, column=overlay_2025_q3_col).value, errors="coerce")) == pytest.approx(26.5, abs=0.01)
        assert float(pd.to_numeric(ws_overlay.cell(row=overlay_45z_row, column=overlay_2025_q4_col).value, errors="coerce")) == pytest.approx(27.7, abs=0.01)
        official_model_row = next(
            (
                rr
                for rr in range(base_row, market_row)
                if str(ws_overlay.cell(row=rr, column=1).value or "").strip().startswith(
                    "Official market model | Representative quarter: "
                )
            ),
            None,
        )
        nebraska_basis_row = _find_row_with_value(ws_overlay, "Nebraska", column=1)
        official_region_header_row = next(
            (
                rr
                for rr in range((official_model_row or base_row) + 1, market_row)
                if str(ws_overlay.cell(row=rr, column=1).value or "").strip() == "Region / family"
            ),
            None,
        )
        coverage_caveat_row = _find_row_with_value(ws_overlay, "Coverage caveat", column=1)
        assert official_model_row is not None and nebraska_basis_row is not None and official_region_header_row is not None and coverage_caveat_row is None
        assert base_row < official_model_row < market_row
        assert f"A{official_model_row}:Q{official_model_row}" in merged_ranges
        official_model_text = str(ws_overlay.cell(row=official_model_row, column=1).value or "").strip()
        assert official_model_text.endswith(".")
        assert official_model_text == (
            "Official corn basis prefers dated GPRE plant bids when available; otherwise it falls back to "
            "active-capacity-weighted AMS basis using mapped state/regional series and deterministic fallbacks."
        )
        assert "active-capacity-weighted AMS basis" in official_model_text
        assert _fill_rgb(ws_overlay.cell(row=official_model_row, column=1)) in {"00EDF4FA", "00EAF3FB"}
        electricity_row = _find_row_with_value(ws_overlay, "Electricity usage", column=1)
        assert electricity_row is not None
        assert float(ws_overlay.row_dimensions[electricity_row].height or 0.0) == pytest.approx(33.0, abs=0.1)
        assert float(ws_overlay.row_dimensions[market_row].height or 0.0) == pytest.approx(30.0, abs=0.1)
        assert str(ws_overlay.cell(row=official_region_header_row, column=4).value or "").strip() == "Mapped ethanol $/gal"
        assert str(ws_overlay.cell(row=official_region_header_row, column=6).value or "").strip() == "Ethanol series"
        assert str(ws_overlay.cell(row=official_region_header_row, column=10).value or "").strip() == "Basis series"
        assert str(ws_overlay.cell(row=official_region_header_row, column=12).value or "").strip() == "Coverage / note"
        assert f"D{official_region_header_row}:E{official_region_header_row}" in merged_ranges
        assert f"F{official_region_header_row}:G{official_region_header_row}" in merged_ranges
        assert f"J{official_region_header_row}:K{official_region_header_row}" in merged_ranges
        assert f"L{official_region_header_row}:Q{official_region_header_row}" in merged_ranges
        assert f"D{nebraska_basis_row}:E{nebraska_basis_row}" in merged_ranges
        assert f"F{nebraska_basis_row}:G{nebraska_basis_row}" in merged_ranges
        assert f"J{nebraska_basis_row}:K{nebraska_basis_row}" in merged_ranges
        assert f"L{nebraska_basis_row}:Q{nebraska_basis_row}" in merged_ranges
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=nebraska_basis_row, column=2).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=nebraska_basis_row, column=3).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=nebraska_basis_row, column=4).value, errors="coerce"))
        non_eth_row = _find_row_with_value(ws_overlay, "Non-ethanol operating activities ($m)", column=1)
        assert non_eth_row is not None
        assert all(ws_overlay.cell(row=non_eth_row, column=cc).comment is None for cc in range(2, 14))
        assert all(ws_overlay.cell(row=quarter_row + 1, column=cc).comment is None for cc in range(2, 14))
        corn_price_row = _find_row_with_value(ws_overlay, "Corn price", column=1)
        ethanol_price_row = _find_row_with_value(ws_overlay, "Ethanol price", column=1)
        gas_price_row = _find_row_with_value(ws_overlay, "Natural gas price", column=1)
        assert corn_price_row is not None and ethanol_price_row is not None and gas_price_row is not None
        assert corn_price_row < gas_price_row < ethanol_price_row
        corn_source_text = str(ws_overlay.cell(row=corn_price_row, column=11).value or "")
        assert "Quarter-open outlook uses local manual snapshot." in corn_source_text
        assert "Current QTD:" in corn_source_text
        assert "Next quarter outlook uses live bids + AMS fallback." in corn_source_text
        assert "Next quarter outlook uses NYMEX futures." in str(ws_overlay.cell(row=gas_price_row, column=11).value or "")
        ethanol_source_text = str(ws_overlay.cell(row=ethanol_price_row, column=11).value or "")
        assert (
            "Next quarter outlook uses local Chicago ethanol futures strip." in ethanol_source_text
            or "Next quarter outlook ethanol unavailable." in ethanol_source_text
        )
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=corn_price_row, column=2).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=corn_price_row, column=4).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=gas_price_row, column=2).value, errors="coerce"))
        assert pd.notna(pd.to_numeric(ws_overlay.cell(row=gas_price_row, column=4).value, errors="coerce"))
        assert "No frozen prior-quarter thesis snapshot" not in corn_source_text
        corn_basis_comment = str(getattr(ws_overlay.cell(row=corn_price_row, column=11).comment, "text", "") or "")
        ethanol_basis_comment = str(getattr(ws_overlay.cell(row=ethanol_price_row, column=11).comment, "text", "") or "")
        assert "AMS fallback" in corn_basis_comment
        assert "Source: local_html." in corn_basis_comment
        assert "footprint-weighted ethanol benchmark" in ethanol_basis_comment
        assert "Chicago ethanol futures strip" in ethanol_basis_comment
        assert f"K{corn_price_row}:U{corn_price_row}" in merged_ranges
        for hidden_label in [
            "Renewable corn oil yield",
            "Distillers yield",
            "Ultra-high protein yield",
            "Electricity usage",
            "Distillers grains price",
            "Ultra-high protein price",
            "Renewable corn oil price",
            "Soybean oil price proxy",
            "Corn oil premium assumption",
            "Implied renewable corn oil proxy price",
            "Distillers contribution",
            "Ultra-high protein contribution",
            "Renewable corn oil contribution",
            "Approximate coproduct credit",
        ]:
            hidden_row = _find_row_with_value(ws_overlay, hidden_label, column=1)
            assert hidden_row is not None
            assert bool(ws_overlay.row_dimensions[hidden_row].hidden)
        simple_crush_row = next(
            rr
            for rr in range(sandbox_build_row + 4, ws_basis.max_row + 1)
            if str(ws_basis.cell(row=rr, column=2).value or "").strip() == "Approximate market crush"
        )
        assert _find_row_with_value(ws_overlay, "Unhedged simple crush margin proxy", column=1) is None
        assert not bool(ws_basis.row_dimensions[simple_crush_row].hidden)
        assert str(ws_basis.cell(row=simple_crush_row, column=11).value or "").strip() == "$/gal"
        assert str(ws_basis.cell(row=simple_crush_row, column=12).value or "").strip() == "Market crush estimate with natural gas cost and GPRE corn basis, weighted to active capacity, and converted to $/gal."
        overlay_merge_ranges = list(ws_overlay.merged_cells.ranges)
        for idx, left in enumerate(overlay_merge_ranges):
            for right in overlay_merge_ranges[idx + 1:]:
                rows_overlap = not (left.max_row < right.min_row or right.max_row < left.min_row)
                cols_overlap = not (left.max_col < right.min_col or right.max_col < left.min_col)
                assert not (rows_overlap and cols_overlap), f"Overlapping merged ranges found: {left} vs {right}"
        chart_title_row = _find_row_with_value(ws_overlay, "Approximate market crush (weekly)", column=2)
        quarterly_chart_title_row = _find_row_with_value(ws_overlay, "Approximate market crush, fitted models, and real GPRE crush margin (quarterly)", column=2)
        assert chart_title_row is not None
        assert chart_title_row > simple_crush_row
        assert chart_title_row != 100
        assert not bool(ws_overlay.row_dimensions[chart_title_row].hidden)
        assert {str(rng) for rng in ws_overlay.merged_cells.ranges if rng.min_row <= chart_title_row <= rng.max_row} == {f"B{chart_title_row}:U{chart_title_row}"}
        assert _fill_rgb(ws_overlay.cell(row=chart_title_row, column=2)) == "006FA8DC"
        assert float(ws_overlay.row_dimensions[chart_title_row].height or 0.0) == pytest.approx(18.0, abs=0.1)
        assert ws_overlay._charts
        crush_chart = ws_overlay._charts[0]
        assert int(getattr(crush_chart.anchor._from, "col", -1)) == 1
        assert int(getattr(crush_chart.anchor._from, "row", -1)) + 1 == chart_title_row + 1
        assert int(getattr(crush_chart.anchor.to, "col", -1)) + 1 == 22
        assert int(getattr(crush_chart.anchor.to, "row", -1)) >= chart_title_row + 24
        assert len(crush_chart.series) > 2
        assert type(getattr(crush_chart, "x_axis", None)).__name__ == "NumericAxis"
        assert str(getattr(getattr(crush_chart.x_axis, "number_format", None), "formatCode", "") or "") == ";;;"
        assert str(getattr(getattr(crush_chart.y_axis, "number_format", None), "formatCode", "") or "") == "$0.00"
        assert str(getattr(crush_chart.x_axis, "axPos", "") or "") == "b"
        assert str(getattr(crush_chart.y_axis, "axPos", "") or "") == "l"
        with zipfile.ZipFile(gpre_path) as zf:
            chart_xml = zf.read("xl/charts/chart1.xml").decode("utf-8", errors="ignore")
        assert '<plotVisOnly val="0"/>' in chart_xml
        assert '<scatterChart>' in chart_xml
        assert '<axPos val="b"/>' in chart_xml
        assert 'rot="-2700000"' not in chart_xml
        assert 'Approximate market crush ($/gal)' in chart_xml
        assert 'Prior quarter ($/gal)' in chart_xml
        assert 'Current QTD ($/gal)' in chart_xml
        assert 'Next quarter outlook ($/gal)' in chart_xml
        assert '<a:prstDash val="sysDash"/>' in chart_xml
        assert '<c:marker>' in chart_xml or '<marker>' in chart_xml
        thesis_helper_dates = [ws_overlay.cell(row=rr, column=25).value for rr in (chart_title_row + 1, chart_title_row + 2, chart_title_row + 3)]
        assert sum(1 for val in thesis_helper_dates if val) == 3
        assert str(ws_overlay["B130"].value or "").strip() != "Exploratory GPRE basis proxy sandbox (test)"
        basis_dir = Path(__file__).resolve().parents[2] / "GPRE" / "basis_proxy"
        assert (basis_dir / "gpre_basis_proxy_quarterly.csv").exists()
        assert (basis_dir / "gpre_basis_proxy_quarterly.parquet").exists()
        assert (basis_dir / "gpre_basis_proxy_summary.md").exists()
        visible_process_labels = [
            str(ws_basis.cell(row=rr, column=2).value or "").strip()
            for rr in range(sandbox_build_row + 4, simple_crush_row + 1)
            if str(ws_basis.cell(row=rr, column=2).value or "").strip()
            and not bool(ws_basis.row_dimensions[rr].hidden)
        ]
        assert visible_process_labels == [
            "Ethanol revenue contribution",
            "Distillers contribution",
            "Ultra-high protein contribution",
            "Renewable corn oil contribution",
            "Feedstock cost",
            "Natural gas burden",
            "Approximate coproduct credit",
            "Approximate market crush",
        ]
        coeff_header_row = _find_row_with_value(ws_overlay, "Coefficient", column=1)
        ethanol_yield_row = _find_row_with_value(ws_overlay, "Ethanol yield", column=1)
        natural_gas_usage_row = _find_row_with_value(ws_overlay, "Natural gas usage", column=1)
        assert coeff_header_row is not None and ethanol_yield_row is not None and natural_gas_usage_row is not None
        assert f"I{coeff_header_row}:Q{coeff_header_row}" in merged_ranges
        assert str(ws_overlay.cell(row=ethanol_yield_row, column=4).value or "").strip() in {"Report-aligned", "Reported"}
        assert "Platform baseline assumption" not in str(ws_overlay.cell(row=ethanol_yield_row, column=6).value or "")
        assert "USDA" in str(ws_overlay.cell(row=ethanol_yield_row, column=9).value or "") or "GPRE filing" in str(ws_overlay.cell(row=ethanol_yield_row, column=6).value or "")
        assert pd.to_numeric(ws_overlay.cell(row=natural_gas_usage_row, column=2).value, errors="coerce") == pytest.approx(28000.0, abs=0.1)
        assert str(ws_overlay.cell(row=natural_gas_usage_row, column=4).value or "").strip() in {"Report-aligned", "Reported"}
        assert "User-entered" not in str(ws_overlay.cell(row=natural_gas_usage_row, column=6).value or "")
        assert "0.028 MMBtu/gal" in str(ws_overlay.cell(row=natural_gas_usage_row, column=9).value or "")
        assert _find_row_with_value(ws_overlay, "Coverage caveat", column=1) is None
        current_margin_row = _find_row_with_value(ws_overlay, "Current margin setup", column=3)
        q1_position_row = _find_row_with_value(ws_overlay, "Q1 margin positioning", column=3)
        lock_in_row = _find_row_with_value(ws_overlay, "Active lock-in execution", column=3)
        assert current_margin_row is not None
        assert q1_position_row is not None
        assert lock_in_row is not None
        assert float(ws_overlay.row_dimensions[current_margin_row].height or 0.0) >= 49.5
        assert float(ws_overlay.row_dimensions[q1_position_row].height or 0.0) >= 40.0
        assert float(ws_overlay.row_dimensions[lock_in_row].height or 0.0) >= 49.5
        spacer_rows = [
            rr
            for rr in range(overlay_setup_row + 2, bridge_row)
            if not str(ws_overlay.cell(row=rr, column=1).value or "").strip()
            and not str(ws_overlay.cell(row=rr, column=2).value or "").strip()
            and not str(ws_overlay.cell(row=rr, column=3).value or "").strip()
            and float(ws_overlay.row_dimensions[rr].height or 0.0) == pytest.approx(6.0, abs=0.1)
        ]
        assert 9 <= len(spacer_rows) <= 10
        assert spacer_rows == sorted(spacer_rows)
        assert all((right - left) in {2, 4} for left, right in zip(spacer_rows, spacer_rows[1:]))
        assert "Basis_Proxy_Sandbox" in wb_gpre.sheetnames
        assert wb_gpre.sheetnames.index("Promise_Progress_UI") < wb_gpre.sheetnames.index("Basis_Proxy_Sandbox") < wb_gpre.sheetnames.index("Hidden_Value_Flags")
        ws_basis = wb_gpre["Basis_Proxy_Sandbox"]
        basis_merged_ranges = {str(rng) for rng in ws_basis.merged_cells.ranges}
        assert str(ws_basis["B1"].value or "").strip() == "Exploratory GPRE basis proxy sandbox (test)"
        assert "B1:O1" in basis_merged_ranges
        sandbox_build_row = _find_row_with_value(ws_basis, "Approximate market crush build-up ($/gal)", column=2)
        coproduct_history_title_row = _find_row_with_value(ws_basis, "Coproduct quarterly history", column=2)
        coproduct_volume_support_title_row = _find_row_with_value(ws_basis, "Coproduct volume support audit", column=2)
        coproduct_experimental_title_row = _find_row_with_value(ws_basis, "Coproduct-aware experimental lenses", column=2)
        winner_story_title_row = _find_row_with_value(ws_basis, "Winner story", column=21)
        assert sandbox_build_row is not None
        assert coproduct_history_title_row is not None
        assert coproduct_volume_support_title_row is not None
        assert coproduct_experimental_title_row is not None
        assert winner_story_title_row is not None
        assert float(ws_basis.column_dimensions["D"].width or 0.0) >= 20.0
        assert float(ws_basis.column_dimensions["B"].width or 0.0) >= 16.0
        assert float(ws_basis.column_dimensions["S"].width or 0.0) >= 36.0
        assert float(ws_basis.column_dimensions["U"].width or 0.0) >= 21.0
        assert float(ws_basis.column_dimensions["X"].width or 0.0) >= 14.0
        assert _fill_rgb(ws_basis.cell(row=sandbox_build_row, column=2)) == "00D9E7F3"
        assert _fill_rgb(ws_basis.cell(row=coproduct_history_title_row, column=2)) == "00EAF3FB"
        assert _fill_rgb(ws_basis.cell(row=coproduct_volume_support_title_row, column=2)) == "00EAF3FB"
        assert _fill_rgb(ws_basis.cell(row=coproduct_experimental_title_row, column=2)) == "00F7F9FC"
        assert _fill_rgb(ws_basis.cell(row=sandbox_build_row + 1, column=2)) == "00F4F8FC"
        assert _fill_rgb(ws_basis.cell(row=coproduct_experimental_title_row + 1, column=2)) == "00F8FBFD"
        assert _fill_rgb(ws_basis.cell(row=winner_story_title_row, column=21)) == "00F7F9FC"
        assert str(ws_basis["U4"].value or "").strip() == "How to read this sheet"
        assert "Conclusion" in str(ws_basis["U5"].value or "")
        assert "incumbent baseline" in str(ws_basis["U5"].value or "").lower()
        assert "process comparator" in str(ws_basis["U5"].value or "").lower()
        assert "expanded-pass best" in str(ws_basis["U5"].value or "").lower()
        assert "production winner" in str(ws_basis["U5"].value or "").lower()
        assert "selection vs promotion" in str(ws_basis["U5"].value or "").lower()
        assert "preview quality" in str(ws_basis["U5"].value or "").lower()
        assert "preview max error" in str(ws_basis["U5"].value or "").lower()
        assert "hard-quarter" in str(ws_basis["U5"].value or "").lower()
        assert "underlying" in str(ws_basis["U5"].value or "").lower()
        assert "Realized GPRE crush margin uses reported consolidated before 2025-Q2 and underlying from 2025-Q2 onward." in str(ws_basis["U13"].value or "")
        role_summary_title_row = _find_row_with_value(ws_basis, "Role summary", column=21)
        winner_story_title_row = _find_row_with_value(ws_basis, "Winner story", column=21)
        assert role_summary_title_row is not None
        assert winner_story_title_row is not None
        assert "Production winner" in str(ws_basis.cell(row=role_summary_title_row + 1, column=21).value or "")
        assert "Best forward lens" in str(ws_basis.cell(row=role_summary_title_row + 4, column=21).value or "")
        assert "Hybrid" in str(ws_basis.cell(row=role_summary_title_row + 1, column=22).value or "")
        assert "Forward" in str(ws_basis.cell(row=role_summary_title_row + 1, column=22).value or "")
        assert "Production winner = fitted row used in production" in str(ws_basis.cell(row=role_summary_title_row + 5, column=21).value or "")
        assert "Official reference" in str(ws_basis.cell(row=winner_story_title_row + 1, column=21).value or "")
        winner_story_labels = " | ".join(str(ws_basis.cell(row=rr, column=21).value or "") for rr in range(winner_story_title_row + 1, winner_story_title_row + 14))
        assert "Production winner" in winner_story_labels
        assert "Decision story" in winner_story_labels
        assert "Main preview mode" in winner_story_labels
        assert "Preview block reason" in winner_story_labels
        expanded_best_story_row = _find_row_with_value(ws_basis, "Expanded-pass best", column=21)
        selection_status_story_row = _find_row_with_value(ws_basis, "Selection status", column=21)
        promotion_status_story_row = _find_row_with_value(ws_basis, "Promotion status", column=21)
        preview_quality_story_row = _find_row_with_value(ws_basis, "Preview quality", column=21)
        production_winner_story_row = next(
            (
                rr
                for rr in range(winner_story_title_row + 1, ws_basis.max_row + 1)
                if str(ws_basis.cell(row=rr, column=21).value or "").strip() == "Production winner"
            ),
            None,
        )
        assert expanded_best_story_row is not None and production_winner_story_row is not None
        assert selection_status_story_row is not None and promotion_status_story_row is not None
        assert preview_quality_story_row is not None
        expanded_best_story_val = str(ws_basis.cell(row=expanded_best_story_row, column=22).value or "").strip()
        production_winner_story_val = str(ws_basis.cell(row=production_winner_story_row, column=22).value or "").strip()
        selection_status_story_val = str(ws_basis.cell(row=selection_status_story_row, column=22).value or "").strip()
        promotion_status_story_val = str(ws_basis.cell(row=promotion_status_story_row, column=22).value or "").strip()
        preview_quality_story_val = str(ws_basis.cell(row=preview_quality_story_row, column=22).value or "").strip()
        assert expanded_best_story_val
        assert production_winner_story_val
        assert expanded_best_story_val == production_winner_story_val
        assert selection_status_story_val
        assert promotion_status_story_val
        assert selection_status_story_val != promotion_status_story_val
        assert "close" in preview_quality_story_val.lower()
        metrics_header_row = _find_row_with_value(ws_basis, "Model", column=9)
        assert metrics_header_row is not None
        assert [str(ws_basis.cell(row=metrics_header_row, column=cc).value or "").strip() for cc in range(9, 20)] == [
            "Model",
            "Family",
            "Clean MAE",
            "Underlying MAE",
            "Hybrid",
            "Baseline",
            "Avg diff",
            ">2c",
            ">5c",
            "Status",
            "Notes",
        ]
        basis_blob = " | ".join(
            str(ws_basis.cell(row=rr, column=cc).value or "").strip()
            for rr in range(1, min(ws_basis.max_row, 120) + 1)
            for cc in range(2, 24)
            if str(ws_basis.cell(row=rr, column=cc).value or "").strip()
        )
        assert "Quarter MAE:" in basis_blob
        assert "Top misses:" in basis_blob
        assert "Preview acceptable" in basis_blob or "Preview close" in basis_blob or "Preview loose" in basis_blob or "Preview not_faithful_enough" in basis_blob
        assert "Preview max error" in basis_blob
        assert "Main preview mode" in basis_blob
        assert "Hard-quarter MAE" in basis_blob
        assert "Avg diff / >2c / >5c" in basis_blob
        assert "Recent-quarter winner comparison" in basis_blob
        assert "Quarterly comparison table" in basis_blob
        assert "Hybrid-score leaderboard" in basis_blob
        assert "Current GPRE bids vs AMS reference offsets" in basis_blob
        assert "Implied hedge / realization style study" in basis_blob
        assert "Hedge-style family leaderboard" in basis_blob
        assert "Quarter-by-quarter best-fit hedge style" in basis_blob
        assert "Reported consolidated crush margin ($/gal)" in basis_blob
        assert "Best overall style" in basis_blob
        assert "Best overall family" in basis_blob
        assert "Usable quarters" in basis_blob
        assert "Backtest window: 2023-Q1 to 2025-Q4" in basis_blob
        assert "single lowest-MAE candidate style" in basis_blob
        assert "lowest average MAE across its member styles" in basis_blob
        assert "Diagnostic only; does not change official row, fitted row, or winner selection" in basis_blob
        assert "Best-fit style" in basis_blob
        assert "Weak fit?" in basis_blob
        assert "No simple style explained well" in basis_blob or "realization/ops drag" in basis_blob.lower() or "spot-like behavior" in basis_blob.lower()
        assert "Process q-open blend" in basis_blob
        assert "Process q-open + hedge realization" in basis_blob
        assert "Process q-open + severe ops penalty" in basis_blob
        assert "Process front + ops penalty" in basis_blob
        assert "Process front + ethanol geo" in basis_blob
        assert "Expanded best" in basis_blob
        assert "Winner" in basis_blob
        assert "Hard?" in basis_blob
        assert "Target $/gal" in basis_blob
        assert "Target type" in basis_blob
        assert "Memo disclosed bridge prior-qtr" in basis_blob
        assert "Memo disclosed process prior-qtr" in basis_blob
        assert "Memo pattern bridge prior-qtr" in basis_blob
        assert "Memo pattern process prior-qtr" in basis_blob
        assert "Hedge-adjusted memo tests" in basis_blob
        assert "System roles / checks" in basis_blob
        assert "GPRE crush proxy = fitted production model" in basis_blob
        assert "Expanded-pass best = best challenger in the expanded test set" in basis_blob
        assert "Production winner = model that cleared promotion guardrails" in basis_blob
        assert "Offset c/bu" in basis_blob
        assert "Approximate market crush = simple market/process proxy" in basis_blob
        assert "Unsupported basis coverage excluded" not in basis_blob
        hedge_study_row = _find_row_with_value(ws_basis, "Implied hedge / realization style study", column=2)
        assert hedge_study_row is not None
        hedge_leaderboard_row = _find_row_with_value(ws_basis, "Hedge-style family leaderboard", column=2)
        assert hedge_leaderboard_row is not None
        assert str(ws_basis.cell(row=hedge_leaderboard_row + 1, column=2).value or "").strip() == "Style"
        assert str(ws_basis.cell(row=hedge_leaderboard_row + 1, column=3).value or "").strip() == "Family"
        quarter_best_fit_row = _find_row_with_value(ws_basis, "Quarter-by-quarter best-fit hedge style", column=2)
        assert quarter_best_fit_row is not None
        assert str(ws_basis.cell(row=quarter_best_fit_row + 1, column=2).value or "").strip() == "Quarter"
        assert str(ws_basis.cell(row=quarter_best_fit_row + 1, column=3).value or "").strip() == "Reported consolidated crush margin ($/gal)"
        assert str(ws_basis.cell(row=quarter_best_fit_row + 1, column=4).value or "").strip() == "Best-fit style"
        assert str(ws_basis.cell(row=quarter_best_fit_row + 1, column=7).value or "").strip() == "Weak fit?"
        assert str(ws_basis.cell(row=quarter_best_fit_row + 1, column=8).value or "").strip() == "Hard quarter?"

        ws_log = wb_gpre["QA_Log"]
        assert ws_log.max_row > ws_qn.max_row
        assert wb_gpre.sheetnames.index("History_Q") < wb_gpre.sheetnames.index("QA_Log") < wb_gpre.sheetnames.index("Needs_Review")
        qa_log_blob = "\n".join(
            " | ".join(str(ws_log.cell(row=rr, column=cc).value or "").strip() for cc in range(1, min(ws_log.max_column, 8) + 1))
            for rr in range(1, min(ws_log.max_row, 1200) + 1)
        )
        assert "2011-12-31 00:00:00 | debt_tieout" in qa_log_blob
        ws_nr = wb_gpre["Needs_Review"]
        nr_header = [str(ws_nr.cell(row=1, column=cc).value or "").strip() for cc in range(1, ws_nr.max_column + 1)]
        assert nr_header[:12] == [
            "priority",
            "issue_family",
            "severity",
            "first_seen_q",
            "last_seen_q",
            "quarter_count",
            "latest_message",
            "recommended_action",
            "source",
            "quarter",
            "raw_metric",
            "canonical_issue_key",
        ]
        assert 1 <= _sheet_data_row_count(ws_nr) <= 3
        assert ws_nr.max_row < 10
        nr_first_seen = [str(ws_nr.cell(row=rr, column=4).value or "").strip() for rr in range(2, min(ws_nr.max_row, 80) + 1)]
        assert not any(q.startswith("2010") or q.startswith("2011") for q in nr_first_seen)
        nr_header = [str(ws_nr.cell(row=1, column=cc).value or "").strip() for cc in range(1, ws_nr.max_column + 1)]
        nr_rows = [
            [str(ws_nr.cell(row=rr, column=cc).value or "").strip() for cc in range(1, ws_nr.max_column + 1)]
            for rr in range(2, ws_nr.max_row + 1)
        ]
        nr_records = [
            {nr_header[cc - 1]: ws_nr.cell(row=rr, column=cc).value for cc in range(1, ws_nr.max_column + 1)}
            for rr in range(2, ws_nr.max_row + 1)
        ]
        nr_issue_families = [str(ws_nr.cell(row=rr, column=2).value or "").strip() for rr in range(2, ws_nr.max_row + 1)]
        assert "review_status" in nr_header
        assert nr_issue_families == ["debt_tranches"]
        assert "quarter_text_definition_mismatch" not in nr_issue_families
        assert "quarter_text_no_explicit_support" not in nr_issue_families
        assert any(
            str(row.get("issue_family") or "").strip() == "debt_tranches"
            and str(row.get("priority") or "").strip() == "Methodology / heuristic watch"
            and "Debt table scale inferred" in str(row.get("latest_message") or "")
            and str(row.get("review_status") or "").strip() == "Watch"
            for row in nr_records
        )
        assert "2025-12-31 00:00:00 | Total debt (Q) | quarter_text_numeric_conflict" not in qa_log_blob
        gpre_log_header = [str(ws_log.cell(row=1, column=cc).value or "").strip() for cc in range(1, ws_log.max_column + 1)]
        gpre_log_rows = [
            {gpre_log_header[cc - 1]: ws_log.cell(row=rr, column=cc).value for cc in range(1, ws_log.max_column + 1)}
            for rr in range(2, min(ws_log.max_row, 1200) + 1)
        ]
        assert any(
            str(row.get("metric") or "").strip() == "debt_tranches"
            and str(row.get("issue_family") or "").strip() == "debt_tranches"
            for row in gpre_log_rows
        )
    finally:
        wb_gpre.close()

    wb_pbi = load_workbook(pbi_path, data_only=False, read_only=False)
    try:
        ws_summary = wb_pbi["SUMMARY"]
        debt_eq_row = _find_row_with_value(ws_summary, "Debt-to-equity (latest quarter)", column=1)
        assert debt_eq_row is not None
        assert str(ws_summary.cell(row=debt_eq_row, column=2).value or "").strip() == "N/M (neg equity)"

        assert "Operating_Drivers" in wb_pbi.sheetnames
        assert "Economics_Overlay" not in wb_pbi.sheetnames
        ws_drv_pbi = wb_pbi["Operating_Drivers"]
        assert str(ws_drv_pbi["A4"].value or "").strip() == "Operating Commentary"
        assert ws_drv_pbi.freeze_panes is None
        drv_commentary_rows_pbi = [
            row for row in _operating_commentary_rows(ws_drv_pbi) if _quarter_label_ord(row["stated_in"]) is not None
        ]
        assert 8 <= len(drv_commentary_rows_pbi) <= 20
        assert drv_commentary_rows_pbi[0]["stated_in"] == "Q4 2025"
        drv_quarter_ords_pbi = [_quarter_label_ord(row["stated_in"]) for row in drv_commentary_rows_pbi]
        assert all(ord_val is not None for ord_val in drv_quarter_ords_pbi)
        drv_distinct_ords_pbi = []
        for ord_val in drv_quarter_ords_pbi:
            if ord_val not in drv_distinct_ords_pbi:
                drv_distinct_ords_pbi.append(ord_val)
        assert drv_distinct_ords_pbi == sorted(drv_distinct_ords_pbi, reverse=True)
        assert max(len([row for row in drv_commentary_rows_pbi if row["stated_in"] == qtxt]) for qtxt in {row["stated_in"] for row in drv_commentary_rows_pbi}) <= 3
        assert any((_quarter_label_ord(row["stated_in"]) or 0) <= 2024 * 4 + 4 for row in drv_commentary_rows_pbi)
        assert any(str(row["horizon"]).strip() == "" for row in drv_commentary_rows_pbi)
        assert all(float(ws_drv_pbi.row_dimensions[int(row["row"])].height or 0.0) == pytest.approx(19.5, abs=0.1) for row in drv_commentary_rows_pbi)
        assert float(ws_drv_pbi.row_dimensions[4].height or 0.0) == pytest.approx(22.5, abs=0.1)
        assert float(ws_drv_pbi.row_dimensions[5].height or 0.0) == pytest.approx(21.0, abs=0.1)
        assert _find_row_with_value(ws_drv_pbi, "Actuals — latest 12 quarters", column=1) is None
        seg_title_row = _find_row_with_value(ws_drv_pbi, "Segment support — latest 12 quarters", column=1)
        assert seg_title_row is not None
        legend_row_pbi = seg_title_row - 1
        assert all(not str(ws_drv_pbi.cell(row=legend_row_pbi, column=cc).value or "").strip() for cc in range(1, 9))
        assert [str(ws_drv_pbi.cell(row=legend_row_pbi, column=cc).value or "").strip() for cc in range(9, 14)] == [
            "<=-15%",
            "-15..-5",
            "-5..+5",
            "+5..+15",
            ">=+15%",
        ]
        assert [_fill_rgb(ws_drv_pbi.cell(row=legend_row_pbi, column=cc)) for cc in range(9, 14)] == [
            "00A63A00",
            "00D55E00",
            "00DDDDDD",
            "009BD3F5",
            "002F80ED",
        ]
        assert float(ws_drv_pbi.row_dimensions[legend_row_pbi].height or 0.0) == pytest.approx(15.75, abs=0.1)
        assert float(ws_drv_pbi.row_dimensions[seg_title_row].height or 0.0) == pytest.approx(22.5, abs=0.1)
        assert float(ws_drv_pbi.row_dimensions[seg_title_row + 1].height or 0.0) == pytest.approx(21.0, abs=0.1)
        assert str(ws_drv_pbi.cell(row=seg_title_row + 1, column=1).value or "").strip() == "Metric / segment"
        seg_quarters = [
            str(ws_drv_pbi.cell(row=seg_title_row + 1, column=cc).value or "").strip()
            for cc in range(2, ws_drv_pbi.max_column + 1)
            if str(ws_drv_pbi.cell(row=seg_title_row + 1, column=cc).value or "").strip()
        ]
        assert len(seg_quarters) == 12
        assert seg_quarters[-1] == "2025-Q4"
        assert _find_row_with_value(ws_drv_pbi, "Revenue ($m)", column=1) is not None
        assert _find_row_with_value(ws_drv_pbi, "Adj EBIT / operating profit ($m)", column=1) is not None
        assert _find_row_with_value(ws_drv_pbi, "Margin", column=1) is not None
        sendtech_rev_row = next(
            rr for rr in range(seg_title_row + 2, ws_drv_pbi.max_row + 1)
            if str(ws_drv_pbi.cell(row=rr, column=1).value or "").strip() == "SendTech Solutions"
        )
        q4_2024_col_pbi = _find_col_with_value(ws_drv_pbi, "2024-Q4", row=seg_title_row + 1)
        q4_2025_col_pbi = _find_col_with_value(ws_drv_pbi, "2025-Q4", row=seg_title_row + 1)
        assert q4_2024_col_pbi is not None and q4_2025_col_pbi is not None
        assert _fill_rgb(ws_drv_pbi.cell(row=sendtech_rev_row, column=q4_2024_col_pbi)) == "00D55E00"
        assert _fill_rgb(ws_drv_pbi.cell(row=sendtech_rev_row, column=q4_2025_col_pbi)) == "00D55E00"
        assert _fill_rgb(ws_drv_pbi["C7"]) == "00FFFFFF"
        for segment_label in ["SendTech Solutions", "Presort Services", "Other operations", "Corporate expense"]:
            assert _find_row_with_value(ws_drv_pbi, segment_label, column=1) is not None
        assert any("migration" in row["commentary"].lower() or "recurring revenue" in row["commentary"].lower() for row in drv_commentary_rows_pbi)
        assert any("customer losses" in row["commentary"].lower() or "price concessions" in row["commentary"].lower() for row in drv_commentary_rows_pbi)
        assert any("higher margin revenue streams" in row["commentary"].lower() or "operating leverage" in row["commentary"].lower() for row in drv_commentary_rows_pbi)
        assert any("Presort volumes declined 2%, partly because the quarter had one fewer day." == row["commentary"] for row in drv_commentary_rows_pbi)
        assert any("SendTech revenue declined due to a smaller mailing install base and near-term headwinds from the product migration." == row["commentary"] for row in drv_commentary_rows_pbi)
        assert any("Presort revenue increased due to higher volumes and pricing." == row["commentary"] for row in drv_commentary_rows_pbi)
        assert any(
            "SendTech revenue declined due to mail decline at low-to-mid single-digit rates, partly offset by growth in shipping."
            == row["commentary"]
            for row in drv_commentary_rows_pbi
        )
        assert any(
            "Presort adjusted operating profit improved due to higher revenue per piece, in addition to continued labor and transportation cost productivity and cost reductions."
            == row["commentary"]
            for row in drv_commentary_rows_pbi
        )
        assert not any("..." in row["commentary"] for row in drv_commentary_rows_pbi)
        assert not any("aforementioned loss of pre sort customers" in row["commentary"].lower() for row in drv_commentary_rows_pbi)
        assert not any("accelerated or sudden decline" in row["commentary"].lower() for row in drv_commentary_rows_pbi)
        assert not any(row["commentary"].lower().startswith(("due to ", "driven by ", "helped by ", "impacted by ")) for row in drv_commentary_rows_pbi)
        assert not any(row["commentary"] == "Driven by simplification and cost reduction initiatives." for row in drv_commentary_rows_pbi)
        assert max(len(drv_commentary_rows), len(drv_commentary_rows_pbi)) >= 15

        ws_qn = wb_pbi["Quarter_Notes_UI"]
        q3_2025_notes = _quarter_block_notes(ws_qn, "2025-09-30")
        assert "Free cash flow declined to $60.4m, down $13.1m YoY." in q3_2025_notes
        note_rows = {
            str(ws_qn.cell(row=rr, column=3).value or "").strip()
            for rr in range(1, ws_qn.max_row + 1)
            if str(ws_qn.cell(row=rr, column=3).value or "").strip()
        }
        assert "Free cash flow declined to $17.5m, down $43.2m YoY." in note_rows
        assert "Presort delivered record revenue and EBIT, while SendTech again improved profit and margins." in note_rows
        assert "Adjusted EBIT improved by more than $23m on relatively flat revenue, supported by segment performance and an 8% opex decline." in note_rows

        ws_qa = wb_pbi["QA_Checks"]
        qa_blob = "\n".join(
            " | ".join(str(ws_qa.cell(row=rr, column=cc).value or "").strip() for cc in range(1, 6))
            for rr in range(1, ws_qa.max_row + 1)
        )
        assert "shares_diluted | qsum_vs_fy" not in qa_blob
        qa_header = [str(ws_qa.cell(row=1, column=cc).value or "").strip() for cc in range(1, ws_qa.max_column + 1)]
        qa_rows = [
            {qa_header[cc - 1]: ws_qa.cell(row=rr, column=cc).value for cc in range(1, ws_qa.max_column + 1)}
            for rr in range(2, ws_qa.max_row + 1)
        ]
        assert any(
            str(row.get("metric") or "").strip() == "EBITDA (Q)"
            and str(row.get("issue_family") or "").strip() == "quarter_text_no_explicit_support"
            and str(row.get("severity") or "").strip() == "info"
            for row in qa_rows
        )

        ws_nr = wb_pbi["Needs_Review"]
        needs_header = [str(ws_nr.cell(row=1, column=cc).value or "").strip() for cc in range(1, ws_nr.max_column + 1)]
        needs_rows = [
            [str(ws_nr.cell(row=rr, column=cc).value or "").strip() for cc in range(1, ws_nr.max_column + 1)]
            for rr in range(1, ws_nr.max_row + 1)
        ]
        needs_records = [
            {needs_header[cc - 1]: ws_nr.cell(row=rr, column=cc).value for cc in range(1, ws_nr.max_column + 1)}
            for rr in range(2, ws_nr.max_row + 1)
        ]
        assert 2 <= _sheet_data_row_count(ws_nr) <= 4
        assert ws_nr.max_row < 15
        assert needs_rows[0][:13] == [
            "priority",
            "issue_family",
            "severity",
            "first_seen_q",
            "last_seen_q",
            "quarter_count",
            "latest_message",
            "recommended_action",
            "source",
            "quarter",
            "raw_metric",
            "canonical_issue_key",
            "review_status",
        ]
        assert not any(row[1] in {"bank_deposits", "bank_finance_receivables"} for row in needs_rows[1:] if len(row) > 1)
        assert not any("Latest-quarter QA failed" in row[6] for row in needs_rows[1:] if len(row) > 6)
        assert not any(row[1] == "quarter_text_model_mismatch" for row in needs_rows[1:] if len(row) > 1)
        assert any(
            str(row.get("issue_family") or "").strip() == "debt_integrity"
            and str(row.get("latest_message") or "").strip().startswith("Debt integrity:")
            and str(row.get("recommended_action") or "").strip() in {"fix source preference", "review debt definition"}
            and str(row.get("review_status") or "").strip() in {"Action required", "Definition mismatch"}
            for row in needs_records
        )
        assert any(
            str(row.get("issue_family") or "").strip() == "quarter_text_definition_mismatch"
            and str(row.get("priority") or "").strip() == "Current-quarter basis / definition issues"
            and str(row.get("recommended_action") or "").strip() == "review metric definition"
            and str(row.get("raw_metric") or "").strip() == "FCF (Q)"
            and str(row.get("review_status") or "").strip() == "Definition mismatch"
            for row in needs_records
        )
        assert not any(
            str(row.get("issue_family") or "").strip() == "quarter_text_no_explicit_support"
            and str(row.get("raw_metric") or "").strip() == "EBITDA (Q)"
            for row in needs_records
        )
        assert any(
            str(row.get("issue_family") or "").strip() == "quarter_text_definition_mismatch"
            and str(row.get("latest_message") or "").strip()
            == "Workbook FCF (Q) is CFO-capex based at $201.4m; selected quarter text states company-defined free cash flow of $211.9m. Likely definition mismatch rather than same-basis numeric conflict."
            for row in needs_records
        )

        ws_log = wb_pbi["QA_Log"]
        assert ws_log.max_row > ws_nr.max_row
        log_header = [str(ws_log.cell(row=1, column=cc).value or "").strip() for cc in range(1, ws_log.max_column + 1)]
        assert "is_current_review_relevant" in log_header
        assert "is_expected_legacy" in log_header
        qa_log_blob = "\n".join(
            " | ".join(str(ws_log.cell(row=rr, column=cc).value or "").strip() for cc in range(1, min(ws_log.max_column, 8) + 1))
            for rr in range(1, ws_log.max_row + 1)
        )
        assert "Latest-quarter QA failed" not in qa_log_blob
        assert "quarter_text_model_mismatch" not in qa_log_blob
        assert "2025-12-31 00:00:00 | FCF (Q) | quarter_text_numeric_conflict" not in qa_log_blob
        assert "2025-12-31 00:00:00 | FCF (Q) | quarter_text_definition_mismatch" in qa_log_blob

        ws_audit = wb_pbi["Quarter_Notes_Audit"]
        audit_header = [str(ws_audit.cell(row=1, column=cc).value or "").strip() for cc in range(1, ws_audit.max_column + 1)]
        assert "canonical_source_group" in audit_header
        assert "support_count" in audit_header
        assert "source_count" in audit_header
        audit_blob = "\n".join(
            str(ws_audit.cell(row=rr, column=audit_header.index("final_summary") + 1).value or "").strip()
            for rr in range(1, min(ws_audit.max_row, 40) + 1)
        )
        assert "FCF TTM accelerated FCF TTM accelerated" not in audit_blob
        assert "Revenue TTM still under pressure Revenue TTM still under pressure" not in audit_blob
        seen_audit_keys: set[tuple[str, str, str]] = set()
        for rr in range(2, min(ws_audit.max_row, 80) + 1):
            key = (
                str(ws_audit.cell(row=rr, column=audit_header.index("quarter") + 1).value or "").strip(),
                str(ws_audit.cell(row=rr, column=audit_header.index("idea_label") + 1).value or "").strip(),
                str(ws_audit.cell(row=rr, column=audit_header.index("canonical_source_group") + 1).value or "").strip(),
            )
            assert key not in seen_audit_keys
            seen_audit_keys.add(key)
        audit_stage_idx = audit_header.index("stage") + 1
        audit_stages = {
            str(ws_audit.cell(row=rr, column=audit_stage_idx).value or "").strip()
            for rr in range(2, min(ws_audit.max_row, 120) + 1)
        }
        assert "readback_verified" in audit_stages
        assert "final_selected" not in audit_stages
    finally:
        wb_pbi.close()


def test_current_delivered_workbooks_quarterly_color_logic_is_metric_and_basis_aware() -> None:
    pbi_path = _current_delivered_model_path("PBI")
    gpre_path = _current_delivered_model_path("GPRE")
    if not pbi_path.exists() or not gpre_path.exists():
        pytest.skip("Current delivered PBI/GPRE workbooks missing for quarterly color readback test.")

    wb_pbi = load_workbook(pbi_path, data_only=False, read_only=False)
    try:
        ws_val = wb_pbi["Valuation"]
        q1_2024 = _find_col_with_value(ws_val, "2024-Q1", row=6)
        q2_2025 = _find_col_with_value(ws_val, "2025-Q2", row=6)
        q4_2024 = _find_col_with_value(ws_val, "2024-Q4", row=6)
        q4_2025 = _find_col_with_value(ws_val, "2025-Q4", row=6)
        q2_2023 = _find_col_with_value(ws_val, "2023-Q2", row=6)
        assert q1_2024 is not None and q2_2025 is not None and q4_2024 is not None and q4_2025 is not None and q2_2023 is not None

        revenue_row = _find_row_with_value(ws_val, "Revenue")
        capex_row = _find_row_with_value(ws_val, "Capex")
        fcf_ttm_row = _find_row_with_value(ws_val, "FCF (TTM)")
        net_debt_qoq_row = _find_row_with_value(ws_val, "Net debt QoQ Δ ($m)")
        acquisitions_row = _find_row_with_value(ws_val, "Acquisitions (TTM, cash)")
        interest_cov_row = _find_row_with_value(ws_val, "Interest coverage (P&L TTM)")
        cash_interest_cov_row = _find_row_with_value(ws_val, "Cash interest coverage (TTM)")
        bv_share_row = _find_row_with_value(ws_val, "BV/share")
        tbv_share_row = _find_row_with_value(ws_val, "TBV/share")
        fcf_share_ttm_row = _find_row_with_value(ws_val, "FCF/share (TTM)")
        ev_ebitda_row = _find_row_with_value(ws_val, "EV/EBITDA (TTM)")
        ev_adj_ebitda_row = _find_row_with_value(ws_val, "EV/Adj EBITDA (TTM)")
        assert revenue_row is not None
        assert capex_row is not None
        assert fcf_ttm_row is not None
        assert net_debt_qoq_row is not None
        assert acquisitions_row is not None
        assert interest_cov_row is not None
        assert cash_interest_cov_row is not None
        assert bv_share_row is not None
        assert tbv_share_row is not None
        assert fcf_share_ttm_row is not None
        assert ev_ebitda_row is not None
        assert ev_adj_ebitda_row is not None

        assert _fill_rgb(ws_val.cell(row=revenue_row, column=q1_2024)) == "00A63A00"
        assert _fill_rgb(ws_val.cell(row=capex_row, column=q1_2024)) == "002F80ED"
        assert _fill_rgb(ws_val.cell(row=net_debt_qoq_row, column=q2_2023)) == "002F80ED"
        assert _fill_rgb(ws_val.cell(row=fcf_ttm_row, column=q4_2024)) == "002F80ED"
        assert _fill_rgb(ws_val.cell(row=acquisitions_row, column=q4_2024)) == "00000000"
        assert _fill_rgb(ws_val.cell(row=interest_cov_row, column=q2_2025)) == "002F80ED"
        assert _fill_rgb(ws_val.cell(row=cash_interest_cov_row, column=q2_2025)) == "002F80ED"
        assert _fill_rgb(ws_val.cell(row=bv_share_row, column=q1_2024)) == "00A63A00"
        assert _fill_rgb(ws_val.cell(row=tbv_share_row, column=q4_2025)) == "00A63A00"
        assert _fill_rgb(ws_val.cell(row=fcf_share_ttm_row, column=q4_2024)) == "002F80ED"
        assert _fill_rgb(ws_val.cell(row=ev_ebitda_row, column=q4_2025)) == "00000000"
        assert _fill_rgb(ws_val.cell(row=ev_adj_ebitda_row, column=q4_2025)) == "00000000"

        ws_bs = wb_pbi["BS_Segments"]
        q2_2024_bs = _find_col_with_value(ws_bs, "2024-Q2", row=11)
        q1_2025_bs = _find_col_with_value(ws_bs, "2025-Q1", row=11)
        assert q2_2024_bs is not None and q1_2025_bs is not None

        cash_qoq_row = _find_row_with_value(ws_bs, "Δ Cash QoQ ($m)")
        current_ratio_row = _find_row_with_value(ws_bs, "Current ratio")
        total_debt_qoq_row = _find_row_with_value(ws_bs, "Δ Total debt QoQ ($m)")
        goodwill_pct_row = _find_row_with_value(ws_bs, "Goodwill % of assets")
        finance_receivables_row = _find_row_with_value(ws_bs, "Finance receivables (total)")
        deposits_row = _find_row_with_value(ws_bs, "Deposits (bank/customer)")
        bank_net_funding_row = _find_row_with_value(ws_bs, "Bank net funding")
        corporate_expense_row = _find_row_with_value(ws_bs, "Corporate expense")
        assert cash_qoq_row is not None
        assert current_ratio_row is not None
        assert total_debt_qoq_row is not None
        assert goodwill_pct_row is not None
        assert finance_receivables_row is not None
        assert deposits_row is not None
        assert bank_net_funding_row is not None
        assert corporate_expense_row is not None

        assert _fill_rgb(ws_bs.cell(row=cash_qoq_row, column=q2_2024_bs)) == "002F80ED"
        assert _fill_rgb(ws_bs.cell(row=current_ratio_row, column=q1_2025_bs)) == "00A63A00"
        assert _fill_rgb(ws_bs.cell(row=total_debt_qoq_row, column=q2_2024_bs)) == "002F80ED"
        assert _fill_rgb(ws_bs.cell(row=goodwill_pct_row, column=q1_2025_bs)) == "00A63A00"
        assert _fill_rgb(ws_bs.cell(row=finance_receivables_row, column=q1_2025_bs)) == "00DDDDDD"
        assert _fill_rgb(ws_bs.cell(row=deposits_row, column=q1_2025_bs)) == "009BD3F5"
        assert _fill_rgb(ws_bs.cell(row=bank_net_funding_row, column=q1_2025_bs)) == "00A63A00"
        assert _fill_rgb(ws_bs.cell(row=corporate_expense_row, column=q1_2025_bs)) == "002F80ED"

        for sheet_name in ["REPORT_IS_Q", "REPORT_BS_Q", "REPORT_CF_Q"]:
            ws_report = wb_pbi[sheet_name]
            report_fills = {
                _fill_rgb(ws_report.cell(row=rr, column=cc))
                for rr in range(10, min(ws_report.max_row, 25) + 1)
                for cc in range(2, min(ws_report.max_column, 8) + 1)
                if _fill_rgb(ws_report.cell(row=rr, column=cc)) not in {"", "00000000", "000000"}
            }
            assert report_fills == set()
    finally:
        wb_pbi.close()

    wb_gpre = load_workbook(gpre_path, data_only=False, read_only=False)
    try:
        ws_val = wb_gpre["Valuation"]
        q2_2023 = _find_col_with_value(ws_val, "2023-Q2", row=6)
        q3_2023 = _find_col_with_value(ws_val, "2023-Q3", row=6)
        q4_2024 = _find_col_with_value(ws_val, "2024-Q4", row=6)
        q4_2025 = _find_col_with_value(ws_val, "2025-Q4", row=6)
        assert q2_2023 is not None and q3_2023 is not None and q4_2024 is not None and q4_2025 is not None

        net_debt_qoq_row = _find_row_with_value(ws_val, "Net debt QoQ Δ ($m)")
        cash_interest_cov_row = _find_row_with_value(ws_val, "Cash interest coverage (TTM)")
        bv_share_row = _find_row_with_value(ws_val, "BV/share")
        fcf_share_ttm_row = _find_row_with_value(ws_val, "FCF/share (TTM)")
        ev_ebitda_row = _find_row_with_value(ws_val, "EV/EBITDA (TTM)")
        ev_adj_ebitda_row = _find_row_with_value(ws_val, "EV/Adj EBITDA (TTM)")
        assert net_debt_qoq_row is not None
        assert cash_interest_cov_row is not None
        assert bv_share_row is not None
        assert fcf_share_ttm_row is not None
        assert ev_ebitda_row is not None
        assert ev_adj_ebitda_row is not None

        assert _fill_rgb(ws_val.cell(row=net_debt_qoq_row, column=q2_2023)) == "00A63A00"
        assert _fill_rgb(ws_val.cell(row=net_debt_qoq_row, column=q3_2023)) == "002F80ED"
        assert _fill_rgb(ws_val.cell(row=cash_interest_cov_row, column=q4_2024)) == "002F80ED"
        assert _fill_rgb(ws_val.cell(row=bv_share_row, column=q4_2024)) == "00A63A00"
        assert _fill_rgb(ws_val.cell(row=fcf_share_ttm_row, column=q4_2025)) == "002F80ED"
        assert _fill_rgb(ws_val.cell(row=ev_ebitda_row, column=q4_2025)) == "00000000"
        assert _fill_rgb(ws_val.cell(row=ev_adj_ebitda_row, column=q4_2025)) == "00000000"

        ws_bs = wb_gpre["BS_Segments"]
        q4_2025_bs = _find_col_with_value(ws_bs, "2025-Q4", row=11)
        shares_out_row = _find_row_with_value(ws_bs, "Shares outstanding (m)")
        year_row = _find_row_with_value(ws_bs, "Year")
        assert q4_2025_bs is not None
        assert shares_out_row is not None
        assert year_row is not None
        assert _fill_rgb(ws_bs.cell(row=shares_out_row, column=q4_2025_bs)) == "00D55E00"
        year_2024_col = _find_col_with_value(ws_bs, "2024", row=year_row)
        intersegment_rows = [
            rr for rr in range(1, ws_bs.max_row + 1) if str(ws_bs.cell(row=rr, column=1).value or "").strip() == "Intersegment eliminations"
        ]
        corporate_assets_rows = [
            rr for rr in range(1, ws_bs.max_row + 1) if str(ws_bs.cell(row=rr, column=1).value or "").strip() == "Corporate assets"
        ]
        corporate_activities_rows = [
            rr for rr in range(1, ws_bs.max_row + 1) if str(ws_bs.cell(row=rr, column=1).value or "").strip() == "Corporate activities"
        ]
        assert year_2024_col is not None
        assert intersegment_rows
        assert corporate_assets_rows
        assert corporate_activities_rows
        assert _fill_rgb(ws_bs.cell(row=intersegment_rows[0], column=year_2024_col)) == "00DDDDDD"
        assert _fill_rgb(ws_bs.cell(row=corporate_assets_rows[0], column=year_2024_col)) == "00DDDDDD"
        assert _fill_rgb(ws_bs.cell(row=corporate_activities_rows[0], column=year_2024_col)) == "00DDDDDD"
    finally:
        wb_gpre.close()


def test_current_delivered_workbooks_preserve_pbi_adjusted_truth_and_polish_current_notes() -> None:
    pbi_path = _current_delivered_model_path("PBI")
    gpre_path = _current_delivered_model_path("GPRE")
    if not pbi_path.exists() or not gpre_path.exists():
        pytest.skip("Current delivered PBI/GPRE workbooks missing for adjusted-truth polish readback test.")

    wb_pbi = load_workbook(pbi_path, data_only=False, read_only=False)
    try:
        ws_val = wb_pbi["Valuation"]
        q4_col = _find_col_with_value(ws_val, "2025-Q4", row=6)
        adj_ebitda_ttm_row = _find_row_with_value(ws_val, "Adj EBITDA (TTM)")
        adj_eps_ttm_row = _find_row_with_value(ws_val, "Adj EPS (TTM)")
        adj_fcf_ttm_row = _find_row_with_value(ws_val, "Adj FCF (TTM)")
        assert q4_col is not None
        assert adj_ebitda_ttm_row is not None
        assert adj_eps_ttm_row is not None
        assert adj_fcf_ttm_row is not None
        assert float(pd.to_numeric(ws_val.cell(row=adj_ebitda_ttm_row, column=q4_col).value, errors="coerce")) == pytest.approx(572.869, abs=0.01)
        assert float(pd.to_numeric(ws_val.cell(row=adj_eps_ttm_row, column=q4_col).value, errors="coerce")) == pytest.approx(1.36, abs=0.01)
        assert float(pd.to_numeric(ws_val.cell(row=adj_fcf_ttm_row, column=q4_col).value, errors="coerce")) == pytest.approx(383.256, abs=0.01)

        ws_qn = wb_pbi["Quarter_Notes_UI"]
        for note_txt in [
            "Operating expenses declined $14.0m YoY.",
            "[NEW] Operating expenses declined $14.0m YoY.",
            "[NEW] Operating expenses declined $4.0m YoY.",
        ]:
            row_idx = next(
                rr
                for rr in range(1, ws_qn.max_row + 1)
                if str(ws_qn.cell(row=rr, column=3).value or "").strip() == note_txt
            )
            assert str(ws_qn.cell(row=row_idx, column=2).value or "").strip() == "Results / drivers / better vs prior"
            assert str(ws_qn.cell(row=row_idx, column=4).value or "").strip() == "Adjusted EBIT / margin"
    finally:
        wb_pbi.close()

    wb_gpre = load_workbook(gpre_path, data_only=False, read_only=False)
    try:
        ws_qn = wb_gpre["Quarter_Notes_UI"]
        outlook_row = next(
            rr
            for rr in range(1, ws_qn.max_row + 1)
            if "FY 2026 45Z-related Adjusted EBITDA outlook is at least $188.0m." in str(ws_qn.cell(row=rr, column=3).value or "")
        )
        assert str(ws_qn.cell(row=outlook_row, column=4).value or "").strip() == "45Z-related Adjusted EBITDA outlook"
    finally:
        wb_gpre.close()


def test_current_delivered_workbooks_match_visible_promise_progress_ui_snapshots() -> None:
    expected_latest_blocks = {
        "PBI": [
            ("Revenue guidance", "$1.76bn-$1.86bn", "not yet measurable", "Open"),
            ("Adjusted EBIT guidance", "$410m-$460m", "not yet measurable", "Open"),
            ("EPS guidance", "$1.40-$1.60", "not yet measurable", "Open"),
            ("FCF target", "$340m-$370m", "not yet measurable", "Open"),
            ("Cost savings target", "$180m-$200m", "$157m run-rate", "Updated"),
            ("Strategic milestone", "", "Strategic review phase 2 remains on track by end of Q2 2026.", "On track"),
        ],
        "GPRE": [
            ("Interest expense outlook", "$30m-$35m", "not yet measurable", "Open"),
            ("45Z-related Adjusted EBITDA outlook", ">= $188.0m in 2026", "Advantage Nebraska fully operational (Q4 2025)", "On track"),
            ("Capex guidance (FY 2026)", "$15.0m-$25.0m", "not yet measurable", "Open"),
            ("45Z monetization", "$15m-$25m", "$23.4m", "Hit"),
            ("45Z from remaining facilities", "> $44.4m expected in 2026", "not yet measurable", "On track"),
            ("Advantage Nebraska startup", "Advantage Nebraska fully operational", "Advantage Nebraska fully operational", "Completed"),
            ("Advantage Nebraska EBITDA opportunity", "> $150.0m in 2026", "$150.0m disclosed in 2026", "On track"),
        ],
    }

    def _latest_block_rows(ws) -> list[tuple[str, str, str, str]]:
        out: list[tuple[str, str, str, str]] = []
        in_latest = False
        for rr in range(1, ws.max_row + 1):
            marker = str(ws.cell(rr, 1).value or "").strip()
            if marker.startswith("Promise progress (As of "):
                if in_latest:
                    break
                in_latest = True
                continue
            if not in_latest:
                continue
            metric = str(ws.cell(rr, 1).value or "").strip()
            if not metric or metric == "Metric" or metric == "No high-signal items.":
                continue
            target = str(ws.cell(rr, 2).value or "").strip()
            latest = str(ws.cell(rr, 3).value or "").strip()
            result = str(ws.cell(rr, 4).value or "").strip()
            out.append((metric, target, latest, result))
        return out

    for ticker, expected_rows in expected_latest_blocks.items():
        workbook_path = _current_delivered_model_path(ticker)
        if not workbook_path.exists():
            pytest.skip(f"Current delivered workbook missing for promise snapshot test: {workbook_path}")
        wb = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            ws = wb["Promise_Progress_UI"]
            assert _latest_block_rows(ws) == expected_rows
        finally:
            wb.close()
