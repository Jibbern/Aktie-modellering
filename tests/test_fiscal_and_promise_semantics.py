import datetime as dt
import hashlib
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pytest
import pandas as pd
from openpyxl import Workbook, load_workbook

from pbi_xbrl.excel_writer_context import (
    _history_q_latest_full_year_actuals_from_workbook,
    _history_q_latest_full_year_period_set,
)
from pbi_xbrl.adjusted_metric_history import (
    build_adjusted_metric_history_selection,
    load_registered_issuer_recast_adjusted_metric_history,
)
from pbi_xbrl.period_resolver import derive_quarter_from_ytd, self_check_period_logic
from tests.workbook_test_resources import delivered_workbook_path, registered_ticker_dir


TICKERS = ("PBI", "GPRE", "ANF")


def _load_model(ticker: str):
    return load_workbook(delivered_workbook_path(ticker, Path(__file__).resolve()), data_only=True, read_only=False)


def _cell_text(value: Any) -> str:
    return str(value or "").strip()


def _history_q_test_workbook(rows: List[Tuple[dt.date, float]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "History_Q"
    ws.append(["quarter", "revenue"])
    for qd, revenue_m in rows:
        ws.append([dt.datetime(qd.year, qd.month, qd.day), revenue_m * 1_000_000])
    return wb


def test_non_calendar_q4_derivation_matches_fy_to_9m_by_period_start() -> None:
    facts = pd.DataFrame(
        [
            {
                "tag": "DepreciationAndAmortization",
                "unit": "USD",
                "val": 141_104_000.0,
                "start_d": dt.date(2023, 1, 29),
                "end_d": dt.date(2024, 2, 3),
                "fp": "FY",
                "fy": 2023,
                "fy_calc": 2025,
                "form": "10-K",
                "filed_d": dt.date(2024, 4, 1),
                "accn": "original-fy-source",
            },
            {
                "tag": "DepreciationAndAmortization",
                "unit": "USD",
                "val": 141_104_000.0,
                "start_d": dt.date(2023, 1, 29),
                "end_d": dt.date(2024, 2, 3),
                "fp": "FY",
                "fy": 2025,
                "fy_calc": 2025,
                "form": "10-K",
                "filed_d": dt.date(2026, 3, 26),
                "accn": "later-comparative",
            },
            {
                "tag": "DepreciationAndAmortization",
                "unit": "USD",
                "val": 105_547_000.0,
                "start_d": dt.date(2023, 1, 29),
                "end_d": dt.date(2023, 10, 28),
                "fp": "Q3",
                "fy": 2023,
                "fy_calc": 2024,
                "form": "10-Q",
                "filed_d": dt.date(2023, 12, 4),
                "accn": "q3-source",
            },
            {
                "tag": "DepreciationAndAmortization",
                "unit": "USD",
                "val": 116_610_000.0,
                "start_d": dt.date(2024, 2, 4),
                "end_d": dt.date(2024, 11, 2),
                "fp": "Q3",
                "fy": 2024,
                "fy_calc": 2025,
                "form": "10-Q",
                "filed_d": dt.date(2024, 12, 6),
                "accn": "wrong-fy-map-if-used",
            },
        ]
    )
    result = derive_quarter_from_ytd(
        facts,
        end=dt.date(2024, 2, 3),
        quarter_index=4,
        fy_fp_to_end={(2025, "FY"): dt.date(2024, 2, 3), (2025, "Q3"): dt.date(2024, 11, 2)},
        prefer_forms=["10-Q", "10-K"],
    )

    assert result is not None
    assert result.source == "derived_ytd_q4"
    assert result.value == pytest.approx(35_557_000.0)
    assert "same FY start" in result.note


def test_non_calendar_q4_ytd_self_check_uses_same_fy_start() -> None:
    facts = pd.DataFrame(
        [
            {
                "tag": "Revenues",
                "unit": "USD",
                "val": 5_266_292_000.0,
                "start_d": dt.date(2025, 2, 2),
                "end_d": dt.date(2026, 1, 31),
                "fp": "FY",
                "fy": 2025,
                "fy_calc": 2026,
                "form": "10-K",
                "filed_d": dt.date(2026, 3, 26),
                "accn": "fy-source",
            },
            {
                "tag": "Revenues",
                "unit": "USD",
                "val": 3_596_490_000.0,
                "start_d": dt.date(2025, 2, 2),
                "end_d": dt.date(2025, 11, 1),
                "fp": "Q3",
                "fy": 2025,
                "fy_calc": 2025,
                "form": "10-Q",
                "filed_d": dt.date(2025, 12, 5),
                "accn": "q3-source",
            },
            {
                "tag": "Revenues",
                "unit": "USD",
                "val": 3_276_490_000.0,
                "start_d": dt.date(2024, 2, 4),
                "end_d": dt.date(2024, 11, 2),
                "fp": "Q3",
                "fy": 2024,
                "fy_calc": 2025,
                "form": "10-Q",
                "filed_d": dt.date(2024, 12, 6),
                "accn": "wrong-map-if-used",
            },
        ]
    )
    audit = pd.DataFrame(
        [
            {
                "quarter": dt.date(2026, 1, 31),
                "metric": "revenue",
                "source": "derived_ytd_q4",
                "filed": dt.date(2026, 3, 26),
            }
        ]
    )

    checks = self_check_period_logic(facts, audit, metric_name="revenue", strictness="ytd")
    ytd_checks = checks[checks["check"] == "ytd_components"]

    assert not ytd_checks.empty
    assert set(ytd_checks["status"]) == {"pass"}


def _sum_history_revenue_m(wb: Any, dates: List[dt.date]) -> float:
    ws = wb["History_Q"]
    headers = {
        re.sub(r"[^a-z0-9]+", "", _cell_text(ws.cell(1, cc).value).lower()): cc
        for cc in range(1, int(ws.max_column or 0) + 1)
    }
    q_col = headers["quarter"]
    revenue_col = headers["revenue"]
    wanted = set(dates)
    total = 0.0
    found: List[dt.date] = []
    for rr in range(2, int(ws.max_row or 0) + 1):
        raw_q = ws.cell(rr, q_col).value
        qd = raw_q.date() if hasattr(raw_q, "date") else raw_q
        if qd in wanted:
            found.append(qd)
            total += float(ws.cell(rr, revenue_col).value or 0.0)
    assert set(found) == wanted, f"History_Q missing revenue rows for {sorted(wanted - set(found))}"
    return total / 1_000_000.0 if abs(total) > 10_000.0 else total


def _section_event(title: str) -> str:
    return re.sub(r"\s+revisions$", "", title.strip(), flags=re.I)


def _promise_revision_blocks(ws: Any) -> Dict[str, List[Tuple[int, Dict[str, Any]]]]:
    blocks: Dict[str, List[Tuple[int, Dict[str, Any]]]] = {}
    rr = 1
    while rr <= int(ws.max_row or 0):
        title = _cell_text(ws.cell(rr, 1).value)
        if not title.endswith("revisions"):
            rr += 1
            continue
        header_row = rr + 1
        headers = {
            _cell_text(ws.cell(header_row, cc).value).lower(): cc
            for cc in range(1, min(int(ws.max_column or 0), 13) + 1)
            if _cell_text(ws.cell(header_row, cc).value)
        }
        rows: List[Tuple[int, Dict[str, Any]]] = []
        body_row = header_row + 1
        while body_row <= int(ws.max_row or 0):
            first = _cell_text(ws.cell(body_row, 1).value)
            if first.endswith("revisions") or first.endswith("guidance progression") or first.endswith("open guidance"):
                break
            values = {name: ws.cell(body_row, col).value for name, col in headers.items()}
            metric = _cell_text(values.get("metric") or values.get("milestone"))
            if metric and metric.lower() not in {"metric", "milestone"}:
                rows.append((body_row, values))
            body_row += 1
        blocks[title] = rows
        rr = body_row
    return blocks


def _all_promise_revision_rows(ticker: str) -> List[Tuple[str, int, Dict[str, Any]]]:
    wb = _load_model(ticker)
    try:
        ws = wb["Promise_Progress_UI"]
        return [
            (block, row_idx, row)
            for block, rows in _promise_revision_blocks(ws).items()
            for row_idx, row in rows
        ]
    finally:
        wb.close()


def _sheet_headers(ws: Any, row_idx: int) -> Dict[str, int]:
    return {
        _cell_text(ws.cell(row_idx, cc).value): cc
        for cc in range(1, int(ws.max_column or 0) + 1)
        if _cell_text(ws.cell(row_idx, cc).value)
    }


def _row_by_label(ws: Any, label: str) -> int:
    for rr in range(1, int(ws.max_row or 0) + 1):
        if _cell_text(ws.cell(rr, 1).value) == label:
            return rr
    raise AssertionError(f"{ws.title}: missing row {label!r}")


def _history_revenue_by_visible_label(wb: Any, ticker: str) -> Dict[str, float]:
    ws = wb["History_Q"]
    headers = {
        re.sub(r"[^a-z0-9]+", "", _cell_text(ws.cell(1, cc).value).lower()): cc
        for cc in range(1, int(ws.max_column or 0) + 1)
    }
    q_col = headers["quarter"]
    rev_col = headers["revenue"]
    fy_col = headers.get("fiscalyear")
    fq_col = headers.get("fiscalquarter")
    out: Dict[str, float] = {}
    for rr in range(2, int(ws.max_row or 0) + 1):
        raw_q = ws.cell(rr, q_col).value
        qd = raw_q.date() if hasattr(raw_q, "date") else raw_q
        if not isinstance(qd, dt.date):
            continue
        revenue = ws.cell(rr, rev_col).value
        if revenue in (None, ""):
            continue
        if ticker == "ANF" and fy_col and fq_col:
            fy = ws.cell(rr, fy_col).value
            fq = ws.cell(rr, fq_col).value
            if fy and fq:
                label = f"{int(fy)}-Q{int(fq)}"
            else:
                label = f"{qd.year}-Q{((qd.month - 1) // 3) + 1}"
        else:
            label = f"{qd.year}-Q{((qd.month - 1) // 3) + 1}"
        out[label] = float(revenue) / 1_000_000.0
    return out


def test_valuation_ttm_uses_hidden_prior_quarters_for_visible_2023() -> None:
    for ticker in ("PBI", "ANF"):
        wb = _load_model(ticker)
        try:
            val = wb["Valuation"]
            quarter_cols = _sheet_headers(val, 6)
            rev_by_label = _history_revenue_by_visible_label(wb, ticker)
            rev_ttm_row = _row_by_label(val, "Revenue (TTM)")
            visible_2023 = [label for label in quarter_cols if re.fullmatch(r"2023-Q[1-4]", label)]
            assert visible_2023, f"{ticker}: Valuation has no visible 2023 quarters"
            checked = 0
            for label in visible_2023:
                year, quarter = map(int, re.match(r"(20\d{2})-Q([1-4])", label).groups())
                labels = []
                y, q = year, quarter
                for _ in range(4):
                    labels.append(f"{y}-Q{q}")
                    q -= 1
                    if q == 0:
                        y -= 1
                        q = 4
                if not all(prev in rev_by_label for prev in labels):
                    continue
                expected = sum(rev_by_label[prev] for prev in labels)
                actual = val.cell(rev_ttm_row, quarter_cols[label]).value
                assert actual not in (None, ""), (
                    f"{ticker} Valuation {label} Revenue (TTM) should use hidden prior quarters {labels}"
                )
                assert float(actual) == pytest.approx(expected, abs=0.002), (
                    f"{ticker} Valuation {label} Revenue (TTM) mismatch"
                )
                checked += 1
            assert checked, f"{ticker}: no visible 2023 quarter had four source-backed revenue quarters to test"
        finally:
            wb.close()


def test_pbi_adj_ebit_ttm_uses_source_backed_hidden_history_for_early_visible_quarters() -> None:
    protected_path = delivered_workbook_path("PBI", Path(__file__).resolve())
    assert hashlib.sha256(protected_path.read_bytes()).hexdigest() == (
        "6482617ad4f412dc5a1e130dc56c72bba5113ddacea5f7d1ab166fad8ddf5689"
    )

    recast = load_registered_issuer_recast_adjusted_metric_history(
        registered_ticker_dir("PBI", Path(__file__).resolve()) / "historical_segment"
    )
    expected_recast_m = {
        pd.Timestamp("2023-03-31"): (68.028, 96.460),
        pd.Timestamp("2023-06-30"): (69.313, 97.313),
        pd.Timestamp("2023-09-30"): (84.044, 112.113),
        pd.Timestamp("2023-12-31"): (86.334, 114.558),
    }
    for period, (ebit_m, ebitda_m) in expected_recast_m.items():
        row = recast[recast["quarter"] == period].iloc[0]
        assert row["adj_ebit"] == pytest.approx(ebit_m * 1_000_000.0)
        assert row["adj_ebitda"] == pytest.approx(ebitda_m * 1_000_000.0)
        assert row["adj_ebit_scope"] == "continuing_operations_current_presentation"
        assert row["adj_ebitda_scope"] == "continuing_operations_current_presentation"
        assert row["adj_ebit_source_occurrence_id"]
        assert row["adj_ebitda_source_occurrence_id"]

    reported_q4 = pd.DataFrame(
        [
            {
                "quarter": pd.Timestamp("2022-12-31"),
                "adj_ebit": 49_267_000.0,
                "adj_ebitda": 88_331_000.0,
                "source": "ex99",
                "confidence": "high",
                "period_type": "quarter",
                "adj_ebit_definition_id": "definition:issuer-reported-consolidated-adjusted-ebit@1",
                "adj_ebitda_definition_id": "definition:issuer-reported-consolidated-adjusted-ebitda@1",
                "adj_ebit_scope": "reported_consolidated_at_period",
                "adj_ebitda_scope": "reported_consolidated_at_period",
                "adj_ebit_source_occurrence_id": "occurrence:pbi:q4-2022:adjusted-ebit",
                "adj_ebitda_source_occurrence_id": "occurrence:pbi:q4-2022:adjusted-ebitda",
            }
        ]
    )
    quarters = pd.to_datetime(
        ["2022-12-31", "2023-03-31", "2023-06-30", "2023-09-30", "2023-12-31"]
    )
    result = build_adjusted_metric_history_selection(
        pd.concat([reported_q4, recast], ignore_index=True, sort=False),
        quarters,
    )
    assert result.ttm_values["adj_ebit"][pd.Timestamp("2023-09-30")] is None
    assert result.ttm_values["adj_ebitda"][pd.Timestamp("2023-09-30")] is None
    assert result.ttm_values["adj_ebit"][pd.Timestamp("2023-12-31")] == pytest.approx(307_719_000.0)
    assert result.ttm_values["adj_ebitda"][pd.Timestamp("2023-12-31")] == pytest.approx(420_444_000.0)


def test_anf_real_workbook_latest_full_year_uses_exact_retail_fiscal_quarters() -> None:
    wb = _load_model("ANF")
    try:
        period_set = _history_q_latest_full_year_period_set(wb, ticker="ANF")
        assert period_set["fiscal_year"] == 2025
        assert period_set["labels"] == ["2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"]
        assert period_set["quarter_dates"] == [
            dt.date(2025, 5, 3),
            dt.date(2025, 8, 2),
            dt.date(2025, 11, 1),
            dt.date(2026, 1, 31),
        ]
        assert period_set["previous_quarter_dates"] == [
            dt.date(2024, 5, 4),
            dt.date(2024, 8, 3),
            dt.date(2024, 11, 2),
            dt.date(2025, 2, 1),
        ]
        assert dt.date(2025, 2, 1) in period_set["previous_quarter_dates"]
        assert dt.date(2025, 2, 1) not in period_set["quarter_dates"]

        fy2025_revenue = _sum_history_revenue_m(wb, period_set["quarter_dates"])
        fy2024_revenue = _sum_history_revenue_m(wb, period_set["previous_quarter_dates"])
        actuals = _history_q_latest_full_year_actuals_from_workbook(wb, ticker="ANF")
        assert fy2025_revenue == pytest.approx(5266.292, abs=0.001)
        assert fy2024_revenue == pytest.approx(4948.587, abs=0.001)
        assert actuals["revenue_m"] == pytest.approx(5266.292, abs=0.001)
        assert actuals["revenue_growth"] == pytest.approx(0.064201, abs=0.0001)
    finally:
        wb.close()


def test_anf_history_and_valuation_include_source_backed_fiscal_q4_duration_metrics() -> None:
    wb = _load_model("ANF")
    try:
        hist = wb["History_Q"]
        headers = {
            re.sub(r"[^a-z0-9]+", "", _cell_text(hist.cell(1, cc).value).lower()): cc
            for cc in range(1, int(hist.max_column or 0) + 1)
        }
        q_col = headers["quarter"]
        da_col = headers["da"]
        ebitda_col = headers["ebitda"]
        interest_col = headers["interestpaid"]
        tax_col = headers["taxpaid"]
        by_quarter: Dict[dt.date, int] = {}
        for rr in range(2, int(hist.max_row or 0) + 1):
            raw_q = hist.cell(rr, q_col).value
            qd = raw_q.date() if hasattr(raw_q, "date") else raw_q
            if isinstance(qd, dt.date):
                by_quarter[qd] = rr

        checks = {
            dt.date(2024, 2, 3): {
                da_col: 35_557_000.0,
                ebitda_col: 258_358_000.0,
                interest_col: 10_726_000.0,
            },
            dt.date(2025, 2, 1): {
                da_col: 37_163_000.0,
                ebitda_col: 293_227_000.0,
                tax_col: 69_419_000.0,
            },
            dt.date(2026, 1, 31): {
                da_col: 40_455_000.0,
                ebitda_col: 276_386_000.0,
                tax_col: 37_051_000.0,
            },
        }
        for qd, expected_by_col in checks.items():
            assert qd in by_quarter, f"ANF History_Q missing fiscal Q4 {qd}"
            rr = by_quarter[qd]
            for col_idx, expected in expected_by_col.items():
                value = hist.cell(rr, col_idx).value
                assert value not in (None, ""), f"ANF History_Q {qd} {hist.cell(1, col_idx).value} is blank"
                assert float(value) == pytest.approx(expected, abs=1.0)

        val = wb["Valuation"]
        quarter_cols = _sheet_headers(val, 6)
        assert "2023-Q4" in quarter_cols and "2024-Q4" in quarter_cols
        for row_label, visible_label in (
            ("EBITDA", "2023-Q4"),
            ("EBITDA (TTM)", "2023-Q4"),
            ("Interest paid", "2023-Q4"),
            ("Tax paid", "2024-Q4"),
        ):
            cell_value = val.cell(_row_by_label(val, row_label), quarter_cols[visible_label]).value
            assert cell_value not in (None, ""), f"ANF Valuation {row_label} {visible_label} should not be blank"
    finally:
        wb.close()


def test_anf_bs_segments_derives_missing_geography_q4_from_annual_source() -> None:
    wb = _load_model("ANF")
    try:
        ws = wb["BS_Segments"]
        quarter_header_row = _row_by_label(ws, "Quarter")
        assert _cell_text(ws.cell(quarter_header_row - 3, 1).value) == "Balance sheet & Segments"
        headers = _sheet_headers(ws, quarter_header_row)
        q4_col = headers.get("2024-Q4")
        assert q4_col is not None

        expected = {
            "Americas": pytest.approx(1319.7, abs=0.2),
            "EMEA": pytest.approx(224.5, abs=0.2),
            "APAC": pytest.approx(40.7, abs=0.2),
        }
        for segment, expected_value in expected.items():
            segment_row = _row_by_label(ws, segment)
            assert ws.cell(segment_row, q4_col).value == expected_value
    finally:
        wb.close()


def test_real_calendar_reporters_keep_calendar_year_quarters() -> None:
    for ticker in ("PBI", "GPRE"):
        wb = _load_model(ticker)
        try:
            period_set = _history_q_latest_full_year_period_set(wb, ticker=ticker)
            assert period_set["fiscal_year"] == 2025, f"{ticker} latest full year"
            assert period_set["labels"] == ["2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"], ticker
            assert period_set["quarter_dates"] == [
                dt.date(2025, 3, 31),
                dt.date(2025, 6, 30),
                dt.date(2025, 9, 30),
                dt.date(2025, 12, 31),
            ], ticker
        finally:
            wb.close()


def test_fiscal_resolver_supports_future_calendar_and_non_calendar_profiles() -> None:
    calendar_wb = _history_q_test_workbook(
        [
            (dt.date(2025, 3, 31), 100),
            (dt.date(2025, 6, 30), 200),
            (dt.date(2025, 9, 30), 300),
            (dt.date(2025, 12, 31), 400),
            (dt.date(2026, 3, 31), 500),
        ]
    )
    calendar_set = _history_q_latest_full_year_period_set(calendar_wb, ticker="FAKECAL")
    assert calendar_set["fiscal_year"] == 2025
    assert calendar_set["labels"] == ["2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"]

    non_calendar_wb = _history_q_test_workbook(
        [
            (dt.date(2025, 5, 3), 100),
            (dt.date(2025, 8, 2), 200),
            (dt.date(2025, 11, 1), 300),
            (dt.date(2026, 1, 31), 400),
            (dt.date(2026, 5, 2), 500),
        ]
    )
    non_calendar_set = _history_q_latest_full_year_period_set(
        non_calendar_wb,
        ticker="FUTURE_RETAIL",
        fiscal_profile={"year_end_month": 1, "year_end_day": 31, "year_label": "start"},
    )
    assert non_calendar_set["fiscal_year"] == 2025
    assert non_calendar_set["quarter_dates"] == [
        dt.date(2025, 5, 3),
        dt.date(2025, 8, 2),
        dt.date(2025, 11, 1),
        dt.date(2026, 1, 31),
    ]


def test_promise_actual_header_is_exact_and_old_header_absent() -> None:
    for ticker in TICKERS:
        wb = _load_model(ticker)
        try:
            ws = wb["Promise_Progress_UI"]
            all_values = [_cell_text(cell.value) for row in ws.iter_rows(min_col=1, max_col=13) for cell in row]
            assert "Actual / latest actual" not in all_values, ticker
            assert "Final actual" not in all_values, ticker
            assert "Actual reported" not in all_values, ticker
            timeline_headers = [
                [_cell_text(ws.cell(rr, cc).value) for cc in range(1, 13)]
                for rr in range(1, int(ws.max_row or 0) + 1)
                if _cell_text(ws.cell(rr, 1).value) == "Metric"
                and _cell_text(ws.cell(rr, 2).value) == "Previous guide"
            ]
            assert timeline_headers, f"{ticker} has no Promise revision timeline header"
            for header in timeline_headers:
                assert header[4] == "Actual", f"{ticker} timeline header uses {header[4]!r}"
                assert header[5] == "Progress / run-rate", f"{ticker} timeline header missing progress column: {header!r}"
                assert header[6] == "Status", f"{ticker} status column shifted incorrectly: {header!r}"
        finally:
            wb.close()


def test_promise_interim_annual_rows_show_quarter_actual_and_ytd_progress() -> None:
    for ticker in TICKERS:
        for block, row_idx, row in _all_promise_revision_rows(ticker):
            metric = _cell_text(row.get("metric") or row.get("milestone"))
            horizon = _cell_text(row.get("horizon"))
            stated = _cell_text(row.get("stated in"))
            actual = _cell_text(row.get("actual"))
            progress = _cell_text(row.get("progress / run-rate"))
            status = _cell_text(row.get("status")).lower()
            if horizon.endswith("year") and re.fullmatch(r"20\d{2}-Q[1-3]", stated):
                assert status not in {"completed", "hit", "missed", "beat"}, (
                    f"{ticker} Promise_Progress_UI!A{row_idx} {metric}: interim annual progress marked final"
                )
                if progress:
                    assert "TTM:" not in progress, (
                        f"{ticker} Promise_Progress_UI!A{row_idx} {metric}: Promise progress should not default to TTM: {progress!r}"
                    )
                    assert re.search(r"YTD:|Run[- ]rate:|progress|operational|qualified|\$", progress, re.I), (
                        f"{ticker} Promise_Progress_UI!A{row_idx} {metric}: progress value lacks basis label: {progress!r}"
                    )
                if metric.lower() in {"revenue guidance", "adjusted ebit guidance", "fcf target", "capex"}:
                    assert actual or progress, (
                        f"{ticker} Promise_Progress_UI!A{row_idx} {metric}: annual guidance row lacks quarter actual/YTD progress"
                    )


def test_promise_revision_rows_are_grouped_by_stated_in_event_and_not_actual_only() -> None:
    allowed_actual_only_notes = re.compile(r"\b(target|plan|expected|milestone|qualified|operational|run[- ]rate|progress)\b", re.I)
    for ticker in TICKERS:
        for block, row_idx, row in _all_promise_revision_rows(ticker):
            metric = _cell_text(row.get("metric") or row.get("milestone"))
            event = _section_event(block)
            stated = _cell_text(row.get("stated in"))
            assert stated == event, f"{ticker} Promise_Progress_UI!A{row_idx} {metric}: stated in {stated!r} outside {block!r}"
            prev = _cell_text(row.get("previous guide"))
            new = _cell_text(row.get("new/current guide"))
            actual = _cell_text(row.get("actual"))
            change = _cell_text(row.get("change type"))
            note = _cell_text(row.get("source / note"))
            if actual and not prev and not new:
                assert allowed_actual_only_notes.search(f"{metric} {change} {note}"), (
                    f"{ticker} Promise_Progress_UI!A{row_idx} {metric}: actual-only row without guide/target/milestone"
                )


def _metric_rank(metric: str) -> int:
    low = metric.lower()
    if any(tok in low for tok in ("revenue", "sales")):
        return 10
    if "operating margin" in low:
        return 20
    if "adjusted ebitda" in low or "adj ebitda" in low or re.search(r"\bebitda\b", low):
        return 30
    if "adjusted ebit" in low or "adj ebit" in low or re.search(r"\bebit\b", low):
        return 40
    if "eps" in low:
        return 50
    if "fcf" in low or "free cash" in low or "cash flow" in low:
        return 60
    if "capex" in low:
        return 70
    if any(tok in low for tok in ("buyback", "repurchase", "share count", "shares", "diluted shares")):
        return 80
    if any(tok in low for tok in ("cost savings", "restructuring", "cash optimization")):
        return 90
    if any(tok in low for tok in ("debt", "leverage", "liquidity")):
        return 100
    if any(tok in low for tok in ("45z", "policy", "facility", "segment", "tariff", "freight", "erp", "marketing")):
        return 110
    return 120


def test_promise_revision_metrics_use_stable_order_inside_blocks() -> None:
    for ticker in TICKERS:
        wb = _load_model(ticker)
        try:
            blocks = _promise_revision_blocks(wb["Promise_Progress_UI"])
            for block, rows in blocks.items():
                ranks = [_metric_rank(_cell_text(row.get("metric") or row.get("milestone"))) for _, row in rows]
                assert ranks == sorted(ranks), f"{ticker} {block}: metric order is unstable: {ranks}"
        finally:
            wb.close()


def test_promise_annual_actuals_are_horizon_matched_not_final_backfills() -> None:
    final_statuses = {"completed", "hit", "missed", "beat"}
    for ticker in TICKERS:
        for block, row_idx, row in _all_promise_revision_rows(ticker):
            metric = _cell_text(row.get("metric") or row.get("milestone"))
            horizon = _cell_text(row.get("horizon"))
            stated = _cell_text(row.get("stated in"))
            actual = _cell_text(row.get("actual"))
            status = _cell_text(row.get("status")).lower()
            if horizon.endswith("year"):
                h_year = int(horizon.split()[0])
                stated_match = re.fullmatch(r"(20\d{2})-Q([1-4])", stated)
                if stated_match:
                    stated_year = int(stated_match.group(1))
                    stated_q = int(stated_match.group(2))
                    if stated_year < h_year:
                        assert not actual, f"{ticker} Promise_Progress_UI!A{row_idx} {metric}: future annual horizon has actual {actual!r}"
                        assert status not in final_statuses, f"{ticker} Promise_Progress_UI!A{row_idx} {metric}: future annual horizon completed early"
                    elif stated_year == h_year and stated_q < 4:
                        assert status not in final_statuses, (
                            f"{ticker} Promise_Progress_UI!A{row_idx} {metric}: interim annual row marked final with actual {actual!r}"
                        )
                if horizon == "2026 year":
                    assert status not in final_statuses, f"{ticker} Promise_Progress_UI!A{row_idx} {metric}: 2026 annual guidance completed before year-end"


def test_gpre_45z_monetization_revision_chain_and_facility_progress() -> None:
    wb = _load_model("GPRE")
    try:
        blocks = _promise_revision_blocks(wb["Promise_Progress_UI"])
        q3_rows = [(idx, row) for idx, row in blocks.get("2025-Q3 revisions", []) if _cell_text(row.get("metric")) == "45Z monetization"]
        q4_rows = [(idx, row) for idx, row in blocks.get("2025-Q4 revisions", []) if _cell_text(row.get("metric")) == "45Z monetization"]
        assert len(q3_rows) == 1, "GPRE 2025-Q3 should contain one initial 45Z monetization row for 2025-Q3 horizon"
        q3_idx, q3_row = q3_rows[0]
        assert _cell_text(q3_row.get("change type")) == "Initial"
        assert _cell_text(q3_row.get("horizon")) == "2025-Q3"
        assert _cell_text(q3_row.get("new/current guide")) == "Q3 45Z value recorded"
        assert "$26.5m" in _cell_text(q3_row.get("actual")), (
            "GPRE 2025-Q3 45Z monetization should show source-disclosed YTD monetization"
        )
        assert "YTD" in _cell_text(q3_row.get("progress / run-rate"))
        assert "$15.0m-$25.0m" not in _cell_text(q3_row.get("new/current guide")), (
            "Q3 actual 45Z value should not be displayed as the separate Q4 monetization guide"
        )
        assert "Q4 guide" in _cell_text(q3_row.get("source / note"))
        hidden_key = _cell_text(wb["Promise_Progress_UI"].cell(q3_idx, 15).value)
        assert "45z_monetization" in hidden_key
        assert "2025_q3" in hidden_key
        assert "2025_q4" not in hidden_key

        audit = wb["Quarter_Notes_Audit"]
        audit_hits = [
            " | ".join(_cell_text(audit.cell(rr, cc).value) for cc in range(1, min(audit.max_column, 10) + 1))
            for rr in range(1, int(audit.max_row or 0) + 1)
            if "2025-09-30" in _cell_text(audit.cell(rr, 1).value)
            and "45Z production tax credits of $26.5 million" in " ".join(
                _cell_text(audit.cell(rr, cc).value) for cc in range(1, min(audit.max_column, 10) + 1)
            )
        ]
        assert audit_hits, "GPRE Q3 45Z source record should remain traceable in Quarter_Notes_Audit"

        narrative = wb["Quarter_Narrative_Data"]
        narrative_hits = [
            rr
            for rr in range(2, int(narrative.max_row or 0) + 1)
            if _cell_text(narrative.cell(rr, 1).value) == "GPRE"
            and _cell_text(narrative.cell(rr, 2).value) == "2025-Q3"
            and _cell_text(narrative.cell(rr, 4).value) == "45Z monetization"
            and "actual $26.5m" in _cell_text(narrative.cell(rr, 5).value)
        ]
        assert narrative_hits, "Quarter_Narrative_Data should carry the Q3 45Z actual semantics"

        assert len(q4_rows) == 1, "GPRE 2025-Q4 should contain one updated/final 45Z monetization row"
        _, q4_row = q4_rows[0]
        assert _cell_text(q4_row.get("change type")) in {"Updated", "Maintained"}
        assert _cell_text(q4_row.get("actual")) == "$23.4m"
        assert _cell_text(q4_row.get("progress / run-rate")) == "YTD: $49.9m"
        assert "$15.0m-$25.0m" in _cell_text(q4_row.get("previous guide"))

        annual_45z = [
            row
            for _, row in blocks.get("2026-Q1 revisions", [])
            if _cell_text(row.get("metric")) == "2026 year 45Z EBITDA guidance"
        ]
        assert annual_45z, "GPRE 2026-Q1 should show 2026 annual 45Z guidance update"
        assert "$55.2m" in _cell_text(annual_45z[0].get("actual"))
        assert "YTD" in _cell_text(annual_45z[0].get("progress / run-rate"))
        assert _cell_text(annual_45z[0].get("status")) != "Completed"

        q1_2026 = [row for _, row in blocks.get("2026-Q1 revisions", []) if _cell_text(row.get("metric")) == "45Z facility qualification"]
        assert q1_2026, "GPRE 2026-Q1 should retain 45Z facility qualification progress row"
        facility = q1_2026[0]
        assert _cell_text(facility.get("actual")) == ""
        assert "All 8" in _cell_text(facility.get("progress / run-rate"))
        assert _cell_text(facility.get("status")) == "On track"
        assert _cell_text(facility.get("status")) != "Completed"
    finally:
        wb.close()


def _valuation_guidance_rows(ws, asof_text: str) -> List[Tuple[str, str, str, str, str]]:
    out: List[Tuple[str, str, str, str, str]] = []
    in_block = False
    for rr in range(1, int(ws.max_row or 0) + 1):
        marker = _cell_text(ws.cell(row=rr, column=15).value)
        if marker.startswith(f"Guidance (As of {asof_text})"):
            in_block = True
            continue
        if in_block and marker.startswith("Guidance (As of "):
            break
        if not in_block:
            continue
        metric = _cell_text(ws.cell(row=rr, column=15).value)
        if metric in {"", "Metric", "A) Updated / mentioned this quarter", "B) Carry-forward", "No guidance items for this quarter."}:
            continue
        out.append(
            (
                metric,
                _cell_text(ws.cell(row=rr, column=17).value),
                _cell_text(ws.cell(row=rr, column=18).value),
                _cell_text(ws.cell(row=rr, column=19).value),
                _cell_text(ws.cell(row=rr, column=27).value),
            )
        )
    return out


def test_gpre_valuation_guidance_uses_curated_conference_rows_not_noisy_transcript_fragments() -> None:
    wb = _load_model("GPRE")
    try:
        rows = _valuation_guidance_rows(wb["Valuation"], "2025-12-31")
        by_key = {(metric, applies): (stated, guidance, trend) for metric, stated, applies, guidance, trend in rows}
        capex = by_key.get(("Capex guidance (2026 year)", "2026 year"))
        assert capex is not None, "GPRE Valuation 2025-Q4 should retain the curated FY2026 capex guidance row"
        assert capex[0] == "2025-Q4"
        assert capex[1] == "$15.0m-$25.0m"

        guidance_blob = "\n".join(guidance for _, _, _, guidance, _ in rows)
        assert "$5million to $10m" not in guidance_blob
        assert "$10.0m" not in guidance_blob
        assert "Excluded from current base; final guidance expected in 2026" in guidance_blob
        assert any("45Z" in metric and applies == "2026 year" for metric, _, applies, _, _ in rows)
    finally:
        wb.close()


def test_pbi_fcf_eps_and_cost_savings_semantics_are_source_specific() -> None:
    wb = _load_model("PBI")
    try:
        ws = wb["Promise_Progress_UI"]
        all_text = "\n".join(_cell_text(cell.value) for row in ws.iter_rows(min_col=1, max_col=10) for cell in row)
        assert "Adjusted EPS / EPS" not in all_text
        blocks = _promise_revision_blocks(ws)
        section_titles = [
            _cell_text(ws.cell(rr, 1).value)
            for rr in range(1, int(ws.max_row or 0) + 1)
            if _cell_text(ws.cell(rr, 1).value).endswith("guidance progression")
        ]
        assert "2024 guidance progression" in section_titles, "PBI should retain a 2024 guidance progression summary"
        assert "2025 guidance progression" in section_titles, "PBI should retain a 2025 guidance progression summary"
        for section in ("2024 guidance progression", "2025 guidance progression"):
            start = section_titles.index(section) if section in section_titles else -1
            assert start >= -1
        progression_metrics = {
            _cell_text(ws.cell(rr, 1).value)
            for rr in range(1, int(ws.max_row or 0) + 1)
            if _cell_text(ws.cell(rr, 1).value) in {"Revenue guidance", "Adjusted EBIT guidance", "Adjusted EPS guidance", "FCF target"}
        }
        assert {"Revenue guidance", "Adjusted EBIT guidance", "Adjusted EPS guidance", "FCF target"}.issubset(
            progression_metrics
        )
        q2_2024_cost = [row for _, row in blocks.get("2024-Q2 revisions", []) if _cell_text(row.get("metric")) == "Cost savings target"]
        assert q2_2024_cost, "PBI 2024-Q2 should show the source-backed initial cost-savings target"
        assert _cell_text(q2_2024_cost[0].get("new/current guide")) == "$120m-$160m"
        assert "$70m" in _cell_text(q2_2024_cost[0].get("actual"))
        assert "$70m" in _cell_text(q2_2024_cost[0].get("progress / run-rate"))
        assert "$157m" not in _cell_text(q2_2024_cost[0].get("progress / run-rate")), "PBI later run-rate was backfilled into 2024-Q2"

        q1_2025_cost = [row for _, row in blocks.get("2025-Q1 revisions", []) if _cell_text(row.get("metric")) == "Cost savings target"]
        assert q1_2025_cost
        assert "$157m" in _cell_text(q1_2025_cost[0].get("actual"))
        assert "$157m" in _cell_text(q1_2025_cost[0].get("progress / run-rate"))
        q1_2026_cost = [row for _, row in blocks.get("2026-Q1 revisions", []) if _cell_text(row.get("metric")) == "Cost savings target"]
        assert q1_2026_cost
        assert "$157m" in _cell_text(q1_2026_cost[0].get("actual"))

        for block, row_idx, row in _all_promise_revision_rows("PBI"):
            metric = _cell_text(row.get("metric"))
            if metric == "FCF target":
                note = _cell_text(row.get("source / note")).lower()
                assert "free cash flow" in note or "fcf" in note, f"PBI Promise_Progress_UI!A{row_idx}: FCF row lacks definition note"
            assert metric != "Adjusted EPS / EPS", f"PBI Promise_Progress_UI!A{row_idx}: generic EPS label remains"

        narrative = wb["Quarter_Narrative_Data"]
        narrative_blob = "\n".join(_cell_text(cell.value) for row in narrative.iter_rows() for cell in row).lower()
        assert "$120m-$160m annual savings target" in narrative_blob
        assert "$70m annualized reductions" in narrative_blob
        assert "$136m" in narrative_blob and "gec" in narrative_blob
        assert "$240m" in narrative_blob and "cash optimization" in narrative_blob
    finally:
        wb.close()


def test_anf_guidance_progression_sections_keep_header_and_rows_together() -> None:
    wb = _load_model("ANF")
    try:
        ws = wb["Promise_Progress_UI"]
        progression_titles: List[Tuple[int, str]] = []
        for rr in range(1, int(ws.max_row or 0) + 1):
            title = _cell_text(ws.cell(rr, 1).value)
            if re.fullmatch(r"20\d{2} guidance progression", title):
                progression_titles.append((rr, title))
        assert progression_titles, "ANF should render guidance progression sections"
        years = [int(title[:4]) for _, title in progression_titles]
        assert years == sorted(years, reverse=True), f"ANF guidance progression sections out of order: {years}"
        for row_idx, title in progression_titles:
            next_title = _cell_text(ws.cell(row_idx + 1, 1).value)
            assert next_title == "Metric", (
                f"ANF {title} at row {row_idx} is detached from its table header; "
                f"next row starts with {next_title!r}"
            )
            first_data = _cell_text(ws.cell(row_idx + 2, 1).value)
            assert first_data and first_data != "Metric" and not first_data.endswith("guidance progression"), (
                f"ANF {title} at row {row_idx} has no immediate source-backed data row"
            )
    finally:
        wb.close()


def test_gpre_milestone_cost_savings_progression_shows_source_backed_actual_progress() -> None:
    wb = _load_model("GPRE")
    try:
        ws = wb["Promise_Progress_UI"]
        start = 0
        for rr in range(1, int(ws.max_row or 0) + 1):
            if _cell_text(ws.cell(rr, 1).value) == "2025 milestone progression":
                start = rr
                break
        assert start, "GPRE milestone progression block missing"
        header = _sheet_headers(ws, start + 1)
        row_idx = 0
        for rr in range(start + 2, int(ws.max_row or 0) + 1):
            first = _cell_text(ws.cell(rr, 1).value)
            if first.endswith("open guidance") or first.endswith("revisions") or first.endswith("guidance progression"):
                break
            if first == "Cost savings target":
                row_idx = rr
                break
        assert row_idx, "GPRE milestone progression should retain cost savings target row"
        actual = _cell_text(ws.cell(row_idx, header["Actual"]).value)
        status = _cell_text(ws.cell(row_idx, header["Status"]).value)
        note = _cell_text(ws.cell(row_idx, header["Notes/source"]).value)
        assert actual, "GPRE milestone cost savings row should show source-backed actual/progress, not blank"
        assert re.search(r"\$|pace|executed|remaining|run[- ]rate|progress", actual, re.I), (
            f"GPRE milestone cost savings actual/progress is not interpretable: {actual!r}"
        )
        assert status in {"On track", "Beat", "Hit"}, f"Unexpected GPRE cost savings status: {status!r}"
        assert note, "GPRE milestone cost savings row should preserve a source note"
    finally:
        wb.close()


def test_pbi_2026_q1_revenue_guidance_update_is_source_backed_and_same_horizon() -> None:
    wb = _load_model("PBI")
    try:
        blocks = _promise_revision_blocks(wb["Promise_Progress_UI"])
        rows = [row for _, row in blocks.get("2026-Q1 revisions", []) if _cell_text(row.get("metric")) == "Revenue guidance"]
        assert rows, "PBI 2026-Q1 should include the source-backed 2026-year Revenue guidance update"
        row = rows[0]
        assert _cell_text(row.get("horizon")) == "2026 year"
        assert _cell_text(row.get("previous guide")) == "$1.76bn-$1.86bn"
        assert _cell_text(row.get("new/current guide")) == "$1.8bn-$1.86bn"
        assert _cell_text(row.get("change type")) in {"Updated", "Raised", "Narrowed"}
        assert "2025 year" not in _cell_text(row.get("previous guide")).lower()
        assert "$477" in _cell_text(row.get("actual")), "PBI 2026-Q1 annual revenue row should show Q1 actual revenue"
        progress = _cell_text(row.get("progress / run-rate"))
        assert progress.startswith("YTD:"), f"PBI 2026-Q1 annual revenue progress should be YTD, got {progress!r}"
        assert "$477" in progress
        assert "TTM:" not in progress

        for metric, actual_fragment, progress_fragment in (
            ("Adjusted EBIT guidance", "$130.4m", "YTD: $130.4m"),
            ("FCF target", "$28.3m", "YTD: $28.3m"),
        ):
            metric_rows = [row for _, row in blocks.get("2026-Q1 revisions", []) if _cell_text(row.get("metric")) == metric]
            assert metric_rows, f"PBI 2026-Q1 should include {metric}"
            metric_row = metric_rows[0]
            assert _cell_text(metric_row.get("horizon")) == "2026 year"
            assert actual_fragment in _cell_text(metric_row.get("actual"))
            assert progress_fragment in _cell_text(metric_row.get("progress / run-rate"))
            assert "TTM:" not in _cell_text(metric_row.get("progress / run-rate"))
    finally:
        wb.close()


def test_pbi_operating_drivers_segment_adj_ebitda_uses_derivable_bs_segment_values() -> None:
    wb = _load_model("PBI")
    try:
        ws = wb["Operating_Drivers"]
        quarter_row = 0
        for rr in range(1, int(ws.max_row or 0) + 1):
            if _cell_text(ws.cell(rr, 1).value) == "Metric / segment":
                prev_title = "\n".join(
                    _cell_text(ws.cell(prior, 1).value)
                    for prior in range(max(1, rr - 4), rr)
                )
                if "Segment support" in prev_title:
                    quarter_row = rr
                    break
        assert quarter_row, "PBI Operating_Drivers segment support table missing"
        quarter_cols = _sheet_headers(ws, quarter_row)
        needed_quarters = {"2024-Q2", "2024-Q4"}
        assert needed_quarters.issubset(quarter_cols), f"PBI Operating_Drivers missing columns {needed_quarters}"

        metric_rows: Dict[str, Dict[str, int]] = {}
        current_metric = ""
        for rr in range(quarter_row + 1, int(ws.max_row or 0) + 1):
            label = _cell_text(ws.cell(rr, 1).value)
            if label == "Actuals — latest 12 quarters":
                break
            if label == "Margin":
                current_metric = ""
                continue
            if label in {"Adj EBIT / operating profit ($m)", "D&A ($m)", "Adj EBITDA ($m)"}:
                current_metric = label
                metric_rows.setdefault(current_metric, {})
                continue
            if current_metric and label:
                metric_rows[current_metric][label] = rr

        for segment in ("SendTech Solutions", "Presort Services"):
            for quarter_label in sorted(needed_quarters):
                col_idx = quarter_cols[quarter_label]
                adj_ebit = ws.cell(metric_rows["Adj EBIT / operating profit ($m)"][segment], col_idx).value
                da = ws.cell(metric_rows["D&A ($m)"][segment], col_idx).value
                adj_ebitda = ws.cell(metric_rows["Adj EBITDA ($m)"][segment], col_idx).value
                assert adj_ebit not in (None, "") and da not in (None, ""), (
                    f"PBI {segment} {quarter_label}: source EBIT/D&A should be available"
                )
                assert adj_ebitda not in (None, ""), (
                    f"PBI Operating_Drivers {segment} {quarter_label}: derivable Adj EBITDA is missing"
                )
                assert float(adj_ebitda) == pytest.approx(float(adj_ebit) + float(da), abs=0.15)
    finally:
        wb.close()


def _bs_segment_metric_rows(ws: Any) -> Dict[str, Dict[str, int]]:
    out: Dict[str, Dict[str, int]] = {}
    current_metric = ""
    in_quarterly = False
    metric_headers = {
        "Revenue",
        "Adjusted EBIT",
        "Segment operating margin %",
        "Depreciation & amortization",
        "Adjusted EBITDA",
    }
    for rr in range(1, int(ws.max_row or 0) + 1):
        label = _cell_text(ws.cell(rr, 1).value)
        if label == "Quarterly segments":
            in_quarterly = True
            current_metric = ""
            continue
        if label == "Annual segments":
            break
        if not in_quarterly:
            continue
        if label in metric_headers:
            current_metric = label
            out.setdefault(current_metric, {})
            continue
        if current_metric and label:
            out.setdefault(current_metric, {})[label] = rr
    return out


def test_pbi_bs_segments_fills_derivable_quarterly_segment_ebitda_and_margins() -> None:
    wb = _load_model("PBI")
    try:
        ws = wb["BS_Segments"]
        quarter_cols = _sheet_headers(ws, 11)
        rows = _bs_segment_metric_rows(ws)
        for segment in ("SendTech Solutions", "Presort Services"):
            for quarter_label, col_idx in quarter_cols.items():
                if not re.fullmatch(r"20\d{2}-Q[1-4]", quarter_label):
                    continue
                revenue = ws.cell(rows["Revenue"][segment], col_idx).value
                adj_ebit = ws.cell(rows["Adjusted EBIT"][segment], col_idx).value
                da = ws.cell(rows["Depreciation & amortization"][segment], col_idx).value
                if revenue not in (None, "") and adj_ebit not in (None, ""):
                    margin = ws.cell(rows["Segment operating margin %"][segment], col_idx).value
                    assert margin not in (None, ""), f"PBI {segment} {quarter_label}: missing segment margin"
                    assert float(margin) == pytest.approx(float(adj_ebit) / float(revenue), abs=0.0005)
                if adj_ebit not in (None, "") and da not in (None, ""):
                    adj_ebitda = ws.cell(rows["Adjusted EBITDA"][segment], col_idx).value
                    assert adj_ebitda not in (None, ""), f"PBI {segment} {quarter_label}: missing derivable adjusted EBITDA"
                    assert float(adj_ebitda) == pytest.approx(float(adj_ebit) + float(da), abs=0.002)
        tiny_margin_issues: List[str] = []
        for segment, row_idx in rows.get("Revenue", {}).items():
            for quarter_label, col_idx in quarter_cols.items():
                if not re.fullmatch(r"20\d{2}-Q[1-4]", quarter_label):
                    continue
                revenue = ws.cell(row_idx, col_idx).value
                margin_row = rows.get("Segment operating margin %", {}).get(segment)
                if margin_row is None or revenue in (None, ""):
                    continue
                try:
                    revenue_f = float(revenue)
                except Exception:
                    continue
                margin = ws.cell(margin_row, col_idx).value
                if revenue_f <= 0 or abs(revenue_f) < 10.0:
                    if margin not in (None, "", "N/A", "N/M"):
                        tiny_margin_issues.append(f"{segment} {quarter_label} revenue={revenue_f} margin={margin}")
                elif margin not in (None, ""):
                    assert abs(float(margin)) < 5.0, f"PBI {segment} {quarter_label}: absurd segment margin {margin}"
        assert not tiny_margin_issues, "Tiny/invalid denominator segment margins should be suppressed: " + "; ".join(tiny_margin_issues)
    finally:
        wb.close()


def test_anf_bs_segments_fills_source_backed_or_derivable_brand_quarters() -> None:
    wb = _load_model("ANF")
    try:
        ws = wb["BS_Segments"]
        quarter_cols = _sheet_headers(ws, 11)
        rows = _bs_segment_metric_rows(ws)
        for segment in ("Hollister", "Abercrombie"):
            assert segment in rows["Revenue"], f"ANF missing {segment} brand revenue row"
        for quarter_label, col_idx in quarter_cols.items():
            if not re.fullmatch(r"20\d{2}-Q[1-4]", quarter_label):
                continue
            hollister = ws.cell(rows["Revenue"]["Hollister"], col_idx).value
            abercrombie = ws.cell(rows["Revenue"]["Abercrombie"], col_idx).value
            total = ws.cell(rows["Revenue"]["Total Company"], col_idx).value
            if total not in (None, ""):
                assert hollister not in (None, ""), f"ANF {quarter_label}: missing Hollister revenue while Total Company exists"
                assert abercrombie not in (None, ""), f"ANF {quarter_label}: missing Abercrombie revenue while Total Company exists"
                assert float(hollister) + float(abercrombie) == pytest.approx(float(total), abs=0.8)
    finally:
        wb.close()


def test_anf_pre_release_can_show_q4_actual_and_fy_progress_for_same_horizon() -> None:
    wb = _load_model("ANF")
    try:
        blocks = _promise_revision_blocks(wb["Promise_Progress_UI"])
        section_titles = {
            _cell_text(wb["Promise_Progress_UI"].cell(rr, 1).value)
            for rr in range(1, int(wb["Promise_Progress_UI"].max_row or 0) + 1)
            if _cell_text(wb["Promise_Progress_UI"].cell(rr, 1).value).endswith(("revisions", "guidance progression"))
        }
        assert any(title.startswith("2024") for title in section_titles), (
            "ANF should retain source-backed 2024 Promise/guidance history when available"
        )
        assert any(title.startswith("2023") or title.startswith("2022") for title in section_titles), (
            "ANF should retain older source-backed Promise/guidance rows from Slides_Guidance"
        )
        pre_rows = blocks.get("2025-Q4 pre-release update revisions", [])
        final_rows = blocks.get("2025-Q4 revisions", [])
        assert pre_rows, "ANF pre-release block missing"
        assert final_rows, "ANF final 2025-Q4 block missing"
        pre_by_metric = {_cell_text(row.get("metric")): row for _, row in pre_rows}
        assert _cell_text(pre_by_metric["Net sales growth"].get("actual")) == "+5.4%"
        assert _cell_text(pre_by_metric["Net sales growth"].get("progress / run-rate")) == "FY: +6%"
        assert "pre-release was issued before final report" in _cell_text(pre_by_metric["Net sales growth"].get("source / note"))
        assert _cell_text(pre_by_metric["Net sales growth"].get("status")) == "On track"
        final_by_metric = {_cell_text(row.get("metric")): row for _, row in final_rows}
        assert _cell_text(final_by_metric["Net sales growth"].get("actual")) == "+5.4%"
        assert _cell_text(final_by_metric["Net sales growth"].get("progress / run-rate")) == "FY: +6%"
        assert _cell_text(final_by_metric["Adjusted EPS"].get("actual")) == "$3.68 adjusted"
        assert _cell_text(final_by_metric["Adjusted EPS"].get("progress / run-rate")) == "FY: $9.86 adjusted"
        assert "Adjusted EPS / EPS" not in "\n".join(
            _cell_text(cell.value) for row in wb["Promise_Progress_UI"].iter_rows(min_col=1, max_col=10) for cell in row
        )
    finally:
        wb.close()


def test_gpre_future_year_guidance_uses_horizon_section_without_duplicate_capex_rows() -> None:
    wb = _load_model("GPRE")
    try:
        blocks = _promise_revision_blocks(wb["Promise_Progress_UI"])
        capex_by_block = {
            block: [(idx, row) for idx, row in rows if _cell_text(row.get("metric")) == "Capex guidance (2025 year)"]
            for block, rows in blocks.items()
        }
        assert not capex_by_block.get("2024-Q4 revisions"), (
            "Future-year 2025 capex guidance should not render as an ordinary 2024-Q4 revision"
        )
        q1_capex = capex_by_block.get("2025-Q1 revisions", [])
        assert len(q1_capex) == 1, "2025 capex guidance should be merged into a single 2025-Q1 timeline row"
        _row_idx, row = q1_capex[0]
        assert _cell_text(row.get("previous guide")) == "$20m-$35m"
        assert _cell_text(row.get("new/current guide")) == "~$20m remaining"
        assert _cell_text(row.get("horizon")) == "2025 year"
        assert _cell_text(row.get("stated in")) == "2025-Q1"
        note = _cell_text(row.get("source / note"))
        assert "2024-Q4" in note and "$20m-$35m" in note
    finally:
        wb.close()
