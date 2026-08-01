from __future__ import annotations

import re

from openpyxl import Workbook
import pytest

import pbi_xbrl.new_ticker_investment_case_formula_surface as formula_surface

from pbi_xbrl.new_ticker_investment_case_formula_surface import (
    CANONICAL_SCENARIOS,
    CANONICAL_VALUATION_MATRIX_HEADERS,
    CANONICAL_VALUATION_METHODS,
    canonical_investment_case_defined_names,
    canonical_valuation_matrix_lookup_expression,
    canonical_valuation_matrix_row,
)
from pbi_xbrl.standard_template_formula_contract import (
    _apply_investment_case_scenario_formulas,
    _apply_scenario_defined_names,
)


ENABLED_FORMULA_IDS = {
    "investment_case_scenario_formulas",
    "investment_case_valuation_matrix_formulas",
    "investment_case_sensitivity_formulas",
}


def _workbook() -> Workbook:
    workbook = Workbook()
    visible = workbook.active
    visible.title = "{ticker}_Investment_Case"
    workbook.create_sheet("{ticker}_Investment_Case_Data")
    _apply_investment_case_scenario_formulas(workbook, ENABLED_FORMULA_IDS)
    _apply_scenario_defined_names(workbook, ENABLED_FORMULA_IDS)
    return workbook


def _matrix_formula_inventory(workbook: Workbook) -> dict[str, str]:
    support = workbook["{ticker}_Investment_Case_Data"]
    return {
        cell.coordinate: str(cell.value)
        for row in support.iter_rows(min_row=2, max_row=25, min_col=57, max_col=66)
        for cell in row
        if cell.data_type == "f"
    }


def _selected_method_consumers(workbook: Workbook) -> dict[str, tuple[str, str]]:
    visible = workbook["{ticker}_Investment_Case"]
    targets: dict[str, str] = {}
    for row, method_id in zip(
        range(121, 126),
        ("pe", "ev_adjusted_ebitda", "ev_revenue", "fcf_yield", "dcf"),
        strict=True,
    ):
        for column in "BCDEFGH":
            targets[f"{column}{row}"] = method_id
    targets.update({coordinate: "blended" for coordinate in ("E126", "F126", "G126")})
    targets.update({coordinate: "dcf" for coordinate in ("H161", "H163", "H165")})
    targets.update({"H193": "pe", "H199": "ev_adjusted_ebitda", "H205": "fcf_yield"})
    assert len(targets) == 44
    return {coordinate: (method_id, str(visible[coordinate].value)) for coordinate, method_id in targets.items()}


def test_canonical_matrix_has_exact_typed_topology_and_formula_ownership() -> None:
    workbook = _workbook()
    support = workbook["{ticker}_Investment_Case_Data"]

    assert tuple(support.cell(1, column).value for column in range(54, 70)) == CANONICAL_VALUATION_MATRIX_HEADERS
    expected_rows = [
        (scenario_label, method_id)
        for scenario_label, _scenario_token, _scenario_column in CANONICAL_SCENARIOS
        for method_id, _name_token, _metric_id, _offset in CANONICAL_VALUATION_METHODS
    ]
    assert [
        (support[f"BB{row}"].value, support[f"BC{row}"].value)
        for row in range(2, 26)
    ] == expected_rows
    assert all(
        support.cell(row, column).protection.locked
        for row in range(2, 26)
        for column in range(54, 70)
    )
    assert len(support.data_validations.dataValidation) == 0
    assert all(support[f"BP{row}"].value == "investment_case_scenario_valuation" for row in range(2, 26))
    assert all(support[f"BQ{row}"].value == "canonical_formula" for row in range(2, 26))


def test_all_scenarios_feed_their_own_methods_and_dcf() -> None:
    workbook = _workbook()
    support = workbook["{ticker}_Investment_Case_Data"]

    for _scenario_label, scenario_token, scenario_column in CANONICAL_SCENARIOS:
        scenario_prefix = "'{ticker}_Investment_Case'!$" + scenario_column + "$"
        for method_id, _name_token, _metric_id, _offset in CANONICAL_VALUATION_METHODS[:-1]:
            row = canonical_valuation_matrix_row(scenario_token, method_id)
            formula = str(support[f"BI{row}"].value)
            assert scenario_prefix in formula or f"BH{row}" in formula
        dcf_row = canonical_valuation_matrix_row(scenario_token, "dcf")
        dcf_formula = str(support[f"BG{dcf_row}"].value)
        assert scenario_prefix + "85" in dcf_formula
        assert scenario_prefix + "88" in dcf_formula
        assert "$B$69" not in dcf_formula
        assert "$H$165" not in dcf_formula

    dcf_formulas = {
        str(support[f"BG{canonical_valuation_matrix_row(token, 'dcf')}"].value)
        for _label, token, _column in CANONICAL_SCENARIOS
    }
    assert len(dcf_formulas) == 4


def test_canonical_economic_formulas_fail_closed_without_error_or_zero_coercion() -> None:
    workbook = _workbook()
    support = workbook["{ticker}_Investment_Case_Data"]

    for row in range(2, 26):
        for column in ("BE", "BF", "BG", "BH", "BI", "BJ", "BL", "BM", "BN"):
            formula = support[f"{column}{row}"].value
            assert isinstance(formula, str) and formula.startswith("=")
            assert "IFERROR(" not in formula
        if support[f"BC{row}"].value != "blended":
            assert "ISNUMBER" in str(support[f"BI{row}"].value)

    for scenario_token in ("Current", "Bear", "Base", "Bull"):
        first = canonical_valuation_matrix_row(scenario_token, "pe")
        last = canonical_valuation_matrix_row(scenario_token, "dcf")
        blend = canonical_valuation_matrix_row(scenario_token, "blended")
        blend_formula = str(support[f"BI{blend}"].value)
        assert "COUNT('{ticker}_Investment_Case'!$B$117:$F$117)" in blend_formula
        assert "COUNTA('{ticker}_Investment_Case'!$B$117:$F$117)" in blend_formula
        assert f"SUM(BM{first}:BM{last})" in blend_formula
        assert f"SUM(BN{first}:BN{last})" in blend_formula
        state_formula = str(support[f"BK{blend}"].value)
        assert '"Invalid method weight"' in state_formula
        assert '"Available-method weights must sum to 100%"' in state_formula


def test_stable_names_are_investment_case_owned_and_keep_fcf_semantics_distinct() -> None:
    workbook = _workbook()
    expected = canonical_investment_case_defined_names()
    actual = {
        name: workbook.defined_names[name].attr_text
        for name in workbook.defined_names
        if str(name).startswith("IC_")
    }

    assert len(expected) == len(actual) == 40
    for name, (sheet_name, coordinate) in expected.items():
        column, row = re.fullmatch(r"([A-Z]+)([0-9]+)", coordinate).groups()
        assert actual[name] == "'" + sheet_name + "'!$" + column + "$" + row
        assert "Valuation" not in actual[name]
        assert "Valuation_Summary" not in actual[name]
        assert "Valuation_Grid" not in actual[name]
    assert actual["IC_Base_FCF_Per_Share"].endswith("!$D$99")
    assert actual["IC_Base_FCF_Yield_Value_Per_Share"].endswith("!$BI$17")
    assert actual["IC_Base_FCF_Per_Share"] != actual["IC_Base_FCF_Yield_Value_Per_Share"]


def test_visible_selected_products_are_projections_of_the_canonical_matrix() -> None:
    workbook = _workbook()
    visible = workbook["{ticker}_Investment_Case"]

    consumers = _selected_method_consumers(workbook)
    for _coordinate, (method_id, formula) in consumers.items():
        assert "COUNTIFS(" in formula
        assert "!$BB$2:$BB$25" in formula
        assert "!$BC$2:$BC$25" in formula
        assert f'"{method_id}"' in formula
        assert "MATCH(1,INDEX(" in formula
        assert not re.search(r"MATCH\([^\n]+\)\+[0-5]\)", formula)
    assert visible["B100"].value == "='{ticker}_Investment_Case_Data'!$BI$7"
    assert visible["C100"].value == "='{ticker}_Investment_Case_Data'!$BI$13"
    assert visible["D100"].value == "='{ticker}_Investment_Case_Data'!$BI$19"
    assert visible["E100"].value == "='{ticker}_Investment_Case_Data'!$BI$25"


def test_two_key_lookup_fails_closed_for_missing_or_duplicate_keys() -> None:
    expression = canonical_valuation_matrix_lookup_expression(
        support_sheet_reference="'Canonical'",
        output_column="BI",
        scenario_expression="$A$1",
        method_id="pe",
    )
    assert expression.startswith("IF(COUNTIFS(")
    assert '")<>1,"",INDEX(' in expression
    assert "MATCH(1,INDEX(" in expression
    assert expression.count("$BB$2:$BB$25") == 2
    assert expression.count("$BC$2:$BC$25") == 2
    with pytest.raises(ValueError, match="Unknown canonical valuation method"):
        canonical_valuation_matrix_lookup_expression(
            support_sheet_reference="'Canonical'",
            output_column="BI",
            scenario_expression="$A$1",
            method_id="missing",
        )


@pytest.mark.parametrize(
    "attribute,reordered",
    (
        ("CANONICAL_VALUATION_METHODS", tuple(reversed(CANONICAL_VALUATION_METHODS))),
        ("CANONICAL_SCENARIOS", tuple(reversed(CANONICAL_SCENARIOS))),
    ),
)
def test_selected_method_resolution_is_independent_of_canonical_tuple_order(
    monkeypatch: pytest.MonkeyPatch,
    attribute: str,
    reordered: tuple,
) -> None:
    baseline = _selected_method_consumers(_workbook())
    monkeypatch.setattr(formula_surface, attribute, reordered)
    reordered_consumers = _selected_method_consumers(_workbook())
    assert reordered_consumers == baseline


def test_invalid_method_weights_are_preserved_and_block_every_scenario_blend() -> None:
    workbook = _workbook()
    support = workbook["{ticker}_Investment_Case_Data"]

    for scenario_token in ("Current", "Bear", "Base", "Bull"):
        for method_id in ("pe", "ev_adjusted_ebitda", "ev_revenue", "fcf_yield", "dcf"):
            row = canonical_valuation_matrix_row(scenario_token, method_id)
            entered = str(support[f"BL{row}"].value)
            effective = str(support[f"BM{row}"].value)
            contribution = str(support[f"BN{row}"].value)
            assert '="",""' in entered
            assert "ISNUMBER" not in entered
            for formula in (effective, contribution):
                assert "COUNT(" in formula and "COUNTA(" in formula
                assert '"<0"' in formula and '">1"' in formula

        blend = canonical_valuation_matrix_row(scenario_token, "blended")
        assert '"Invalid method weight"' in str(support[f"BK{blend}"].value)
        for column in ("BI", "BM", "BN"):
            formula = str(support[f"{column}{blend}"].value)
            assert "COUNT(" in formula and "COUNTA(" in formula
            assert '"<0"' in formula and '">1"' in formula


def test_sensitivity_axes_use_canonical_total_metric_centers() -> None:
    workbook = _workbook()
    visible = workbook["{ticker}_Investment_Case"]

    for coordinate in ("A201", "A202", "A203"):
        formula = str(visible[coordinate].value)
        assert '"ev_adjusted_ebitda"' in formula
        assert "!$BE$2:$BE$25" in formula
        assert "$D$95" not in formula
    for coordinate in ("A207", "A208", "A209"):
        formula = str(visible[coordinate].value)
        assert '"fcf_yield"' in formula
        assert "!$BE$2:$BE$25" in formula
        assert "$D$99" not in formula
    assert '"ev_adjusted_ebitda"' in str(visible["H199"].value)
    assert '"fcf_yield"' in str(visible["H205"].value)


def test_formula_serialization_is_deterministic() -> None:
    first = _workbook()
    second = _workbook()
    assert _matrix_formula_inventory(first) == _matrix_formula_inventory(second)
