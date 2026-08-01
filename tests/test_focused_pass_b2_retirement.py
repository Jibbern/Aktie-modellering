from __future__ import annotations

import ast
import json
from pathlib import Path
import re
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import range_boundaries

from pbi_xbrl.new_ticker_investment_case_formula_surface import (
    CANONICAL_SCENARIOS,
    CANONICAL_VALUATION_METHODS,
    canonical_investment_case_defined_names,
)
from pbi_xbrl.standard_template_formula_contract import (
    USER_INPUT_CONTRACTS,
    formula_target_contracts,
)
from scripts.materialize_standard_template_shell import _prune_unused_differential_styles


ROOT = Path(__file__).resolve().parents[1]
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
MODULE_MANIFEST = ROOT / "docs" / "workbook_module_manifest.json"
SHELL_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
STYLE_POLICY = ROOT / "docs" / "standard_template_style_policy.json"
NORMALIZED_SCHEMA = ROOT / "docs" / "normalized_company_data.schema.json"
MINIMAL_PACKAGE = ROOT / "tests" / "fixtures" / "normalized_packages" / "TEST_minimal_valid.json"
SHELL_VALIDATOR = ROOT / "scripts" / "validate_standard_template_shell.py"
VISUAL_GAP_AUDIT = ROOT / "docs" / "standard_template_shell_visual_gap_audit.json"

RETIRED_SHEETS = {"Valuation_Summary", "Valuation_Grid"}
RETIRED_BINDINGS = {
    "valuation_output_rows",
    "valuation_input_as_of",
    "valuation_input_shares_outstanding",
    "valuation_input_diluted_shares",
    "valuation_input_net_debt",
    "valuation_input_base_ebitda_ttm",
    "valuation_input_adjusted_ebitda_ttm",
    "valuation_input_fcf_ttm",
    "valuation_input_operating_cash_flow_ttm",
    "valuation_input_revenue_ttm",
    "valuation_input_eps_ttm",
    "valuation_input_adjusted_eps_ttm",
    "valuation_input_book_value_per_share",
    "valuation_input_tangible_book_value_per_share",
    "valuation_input_capex_ttm",
    "valuation_input_interest_paid_ttm",
    "valuation_input_net_income_ttm",
}
RETIRED_FORMULA_IDS = {
    "valuation_output_formulas",
    "valuation_sidecar_formulas",
    "valuation_scenario_formulas",
    "scenario_revenue_route_formulas",
    "valuation_summary_formulas",
    "valuation_grid_formulas",
}
RETIRED_DEFINED_NAMES = {
    "Adj_EBITDA",
    "Adj_EPS_TTM",
    "Adj_FCF_TTM",
    "AsOfQuarter",
    "AutoImpliedGT",
    "Base_EBITDA",
    "BV_PerShare",
    "Capex_TTM",
    "DCF_EV",
    "DCF_FCFF",
    "DCF_Growth",
    "DCF_Horizon",
    "DCF_ImpliedPrice",
    "DCF_TerminalGrowth",
    "DCF_WACC",
    "EPS_TTM",
    "EqShare_Target_Adj",
    "EqShare_Target_EV",
    "EqShare_Target_Yield",
    "Equity_FCF_Yield",
    "EV",
    "FCF_TTM",
    "FCFF_Proxy_TTM",
    "Implied_EV_AdjEBITDA",
    "Implied_EV_EBITDA",
    "Implied_FCFF_Yield",
    "ImpliedGT_Output",
    "ImpliedGT_Status",
    "ImpliedGT_WACC",
    "InterestPaid_TTM",
    "MaintCapexRatio",
    "MarketCap",
    "NetDebt",
    "NetIncome_TTM",
    "OwnerEarnings_TTM",
    "OwnerEarnings_Yield",
    "PerShareMode",
    "Price",
    "RecurringCashCosts",
    "ResolvedRevenueGrowth_Base",
    "ResolvedRevenueGrowth_Bear",
    "ResolvedRevenueGrowth_Bull",
    "ResolvedRevenueGrowth_Custom",
    "Revenue_TTM",
    "ScenarioAdjustedEBITDA",
    "ScenarioAdjustedMargin",
    "ScenarioBaseEBITDA",
    "ScenarioBaseMargin",
    "ScenarioBuybackCash",
    "ScenarioBuybackPrice",
    "ScenarioCapexChange",
    "ScenarioCashInterestChange",
    "ScenarioDebtPaydown",
    "ScenarioEPS",
    "ScenarioEquityValue",
    "ScenarioEV",
    "ScenarioEVAdjustedEBITDA",
    "ScenarioEVBaseEBITDA",
    "ScenarioEVRevenue",
    "ScenarioFCF",
    "ScenarioFCFYield",
    "ScenarioGrowth",
    "ScenarioHorizon",
    "ScenarioImpliedPrice",
    "ScenarioInterestTaxTreatment",
    "ScenarioNetDebt",
    "ScenarioPE",
    "ScenarioPreTaxBridge",
    "ScenarioProfile",
    "ScenarioRevenue",
    "ScenarioShareIssuance",
    "ScenarioShares",
    "ScenarioTaxRate",
    "ScenarioTaxTreatment",
    "ScenarioUpside",
    "ScenarioWCAdjustment",
    "Shares",
    "SharesDiluted",
    "Target_EV_AdjEBITDA",
    "Target_EV_EBITDA",
    "Target_EV_Revenue",
    "Target_EV_Yield",
    "Target_PE",
    "TBV_PerShare",
    "valuation_input_adjusted_ebitda_ttm",
    "valuation_input_adjusted_eps_ttm",
    "valuation_input_as_of",
    "valuation_input_base_ebitda_ttm",
    "valuation_input_book_value_per_share",
    "valuation_input_capex_ttm",
    "valuation_input_diluted_shares",
    "valuation_input_eps_ttm",
    "valuation_input_fcf_ttm",
    "valuation_input_interest_paid_ttm",
    "valuation_input_net_debt",
    "valuation_input_net_income_ttm",
    "valuation_input_operating_cash_flow_ttm",
    "valuation_input_revenue_ttm",
    "valuation_input_shares_outstanding",
    "valuation_input_tangible_book_value_per_share",
    "WCNormalization",
}
RETIRED_EMPTY_RANGES = (
    "O63:X75",
    "G192:AA198",
    "A199:AA261",
    "N262:S271",
)
LEGACY_WRITER_MODULES = {
    "pbi_xbrl.valuation",
    "pbi_xbrl.excel_writer",
    "pbi_xbrl.excel_writer_context",
    "pbi_xbrl.excel_writer_financials",
    "pbi_xbrl.excel_writer_valuation",
    "pbi_xbrl.excel_writer_valuation_orchestrator",
    "pbi_xbrl.excel_writer_valuation_formula_core_render",
    "pbi_xbrl.excel_writer_valuation_final_layout",
    "pbi_xbrl.excel_writer_valuation_operating_thesis_render",
    "pbi_xbrl.excel_writer_valuation_sensitivity_heatmap_render",
    "pbi_xbrl.excel_writer_valuation_trend_flags_render",
    "pbi_xbrl.excel_writer_valuation_history_grid_render",
    "pbi_xbrl.excel_writer_valuation_guidance_render",
    "pbi_xbrl.excel_writer_valuation_hidden_value_render",
    "pbi_xbrl.excel_writer_valuation_debt_detail_render",
    "pbi_xbrl.excel_writer_valuation_precompute",
    "pbi_xbrl.excel_writer_anf_investment_case",
    "pbi_xbrl.excel_writer_sector_investment_case",
    "pbi_xbrl.excel_writer_investment_case_support",
    "pbi_xbrl.excel_writer_post_quarter_capital_events",
}


def _json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _range_cells(worksheet, target: str):
    min_col, min_row, max_col, max_row = range_boundaries(target)
    return (
        cell
        for row in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        )
        for cell in row
    )


def _ranges_intersect(left: str, right: str) -> bool:
    left_min_col, left_min_row, left_max_col, left_max_row = range_boundaries(left)
    right_min_col, right_min_row, right_max_col, right_max_row = range_boundaries(right)
    return not (
        left_max_col < right_min_col
        or right_max_col < left_min_col
        or left_max_row < right_min_row
        or right_max_row < left_min_row
    )


def test_retired_engine_is_physically_absent_from_the_checked_in_shell() -> None:
    workbook = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        assert RETIRED_SHEETS.isdisjoint(workbook.sheetnames)
        assert len(RETIRED_DEFINED_NAMES) == 101
        assert RETIRED_DEFINED_NAMES.isdisjoint(workbook.defined_names)

        valuation = workbook["Valuation"]
        for target in RETIRED_EMPTY_RANGES:
            cells = tuple(_range_cells(valuation, target))
            assert all(cell.value is None for cell in cells), target
            assert all(cell.protection.locked for cell in cells), target
            assert all(cell.comment is None and cell.hyperlink is None for cell in cells), target
            assert all(cell.style_id == 0 for cell in cells), target

        assert {str(cell_range) for validation in valuation.data_validations.dataValidation
                for cell_range in validation.ranges.ranges
                if any(cell.coordinate in validation.cells for cell in _range_cells(valuation, "A192:AA271"))} == set()
        assert not any(
            any(cell.coordinate in merged for cell in _range_cells(valuation, target))
            for target in RETIRED_EMPTY_RANGES
            for merged in valuation.merged_cells.ranges
        )
        assert not any(
            _ranges_intersect(str(cell_range), target)
            for conditional_range in valuation.conditional_formatting
            for cell_range in conditional_range.sqref.ranges
            for target in RETIRED_EMPTY_RANGES
        )

        all_formulas = [
            str(cell.value)
            for sheet in workbook.worksheets
            for cell in sheet._cells.values()
            if cell.data_type == "f"
        ]
        assert not any("Valuation_Summary" in formula or "Valuation_Grid" in formula for formula in all_formulas)
        retired_reference = re.compile(r"'?Valuation'?\!\$?[A-Z]+\$?(?:19[2-9]|2[0-5][0-9]|26[01])\b")
        assert not any(retired_reference.search(formula) for formula in all_formulas)
    finally:
        workbook.close()


def test_differential_style_pruning_is_live_reference_driven() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Example"
    worksheet["A1"].fill = PatternFill("solid", fgColor="F4B183")
    for color in ("F4B183", "F8CBAD"):
        workbook._differential_styles.add(  # type: ignore[attr-defined]
            DifferentialStyle(fill=PatternFill("solid", fgColor=color))
        )
    live_fill = PatternFill("solid", fgColor="F2F4F5")
    worksheet.conditional_formatting.add(
        "B1:B2",
        FormulaRule(formula=["B1>0"], fill=live_fill),
    )
    worksheet.conditional_formatting.add(
        "C1:C2",
        FormulaRule(formula=["C1>0"], fill=live_fill),
    )

    result = _prune_unused_differential_styles(workbook)

    assert result == {"before_count": 2, "after_count": 1, "live_rule_count": 2}
    assert len(workbook._differential_styles.styles) == 1  # type: ignore[attr-defined]
    assert all(
        rule.dxfId == 0
        for rules in worksheet.conditional_formatting._cf_rules.values()
        for rule in rules
    )
    assert worksheet["A1"].fill.fgColor.rgb == "00F4B183"
    assert workbook._differential_styles.styles[0].fill.fgColor.rgb == "00F2F4F5"  # type: ignore[attr-defined]
    workbook.close()


def test_checked_in_shell_has_no_orphan_or_obsolete_differential_styles() -> None:
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with ZipFile(SHELL) as archive:
        styles = ElementTree.fromstring(archive.read("xl/styles.xml"))
        dxfs_node = styles.find(f"{namespace}dxfs")
        assert dxfs_node is not None
        dxfs = list(dxfs_node)
        referenced_ids = {
            int(rule.attrib["dxfId"])
            for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
            for rule in ElementTree.fromstring(archive.read(name)).iter(f"{namespace}cfRule")
            if "dxfId" in rule.attrib
        }
        dxf_colors = {
            color.attrib["rgb"][-6:].upper()
            for dxf in dxfs
            for color in dxf.iter(f"{namespace}fgColor")
            if "rgb" in color.attrib
        }

    assert referenced_ids == set(range(len(dxfs)))
    assert dxf_colors.isdisjoint({"F4B183", "F8CBAD"})
    assert len(dxfs) == 1

    workbook = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        investment_case = workbook["{ticker}_Investment_Case"]
        rules = {
            str(conditional_range.sqref): tuple(
                (rule.type, rule.priority, tuple(rule.formula), rule.dxf.fill.fgColor.rgb)
                for rule in conditional_rules
            )
            for conditional_range, conditional_rules in investment_case.conditional_formatting._cf_rules.items()
        }
        assert rules == {
            "A47:L48": (("expression", 1, ('LOWER($B$42)<>LOWER("Brand")',), "00F2F4F5"),),
            "A50:L52": (("expression", 2, ('LOWER($B$42)<>LOWER("Geography")',), "00F2F4F5"),),
            "A45:L45": (("expression", 3, ('LOWER($B$42)<>LOWER("Total Company")',), "00F2F4F5"),),
        }
        assert len(workbook["Valuation"].conditional_formatting) == 0
    finally:
        workbook.close()


def test_retired_sidecar_has_no_active_validator_or_visual_gap_expectation() -> None:
    validator_source = SHELL_VALIDATOR.read_text(encoding="utf-8")
    for retired_expectation in (
        '"O63": "Output"',
        '"U63": "Value"',
        '"X63": "Interpretation"',
    ):
        assert retired_expectation not in validator_source

    audit = _json(VISUAL_GAP_AUDIT)
    issues = list(audit["shell_validation"].get("issues") or [])
    assert not any(
        issue.get("rule_id") == "valuation_guidance_sidecar_header_missing"
        and issue.get("target") in {"O63", "U63", "X63"}
        for issue in issues
    )
    assert not any(
        issue.get("rule_id") == "rich_shell_merge_family_sparse"
        and issue.get("severity") in {"P0", "P1"}
        for issue in issues
    )


def test_compact_valuation_summary_is_locked_and_only_references_canonical_names() -> None:
    workbook = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        valuation = workbook["Valuation"]
        assert valuation["A192"].value == "Forward Valuation Summary"
        assert tuple(valuation.cell(193, column).value for column in range(1, 7)) == (
            "Metric",
            "Current baseline",
            "Bear",
            "Base",
            "Bull",
            "State / context",
        )
        metric_tokens = (
            "GAAP_EPS",
            "Adjusted_EBITDA",
            "FCF_Per_Share",
            "Blended_Value_Per_Share",
            "Upside_Downside",
        )
        scenarios = ("Current", "Bear", "Base", "Bull")
        for row, metric in zip(range(194, 199), metric_tokens, strict=True):
            assert valuation.row_dimensions[row].height == 36
            assert valuation.cell(row, 6).alignment.wrap_text
            for column, scenario in zip(range(2, 6), scenarios, strict=True):
                cell = valuation.cell(row, column)
                assert cell.value == f"=IC_{scenario}_{metric}"
                assert cell.protection.locked
        assert not any(
            cell.protection.locked is False
            for cell in _range_cells(valuation, "A192:F198")
        )
        assert not any(
            any(cell.coordinate in validation.cells for cell in _range_cells(valuation, "A192:F198"))
            for validation in valuation.data_validations.dataValidation
        )
    finally:
        workbook.close()


def test_old_bindings_formula_ids_inputs_and_normalized_outputs_are_deleted() -> None:
    bindings = _json(BINDING_MAP)
    binding_ids = {str(row["binding_id"]) for row in bindings["bindings"]}
    assert RETIRED_BINDINGS.isdisjoint(binding_ids)

    formula_ids = {contract.formula_id for contract in formula_target_contracts()}
    assert RETIRED_FORMULA_IDS.isdisjoint(formula_ids)
    module_payload = _json(MODULE_MANIFEST)
    module_formula_ids = {
        str(formula_id)
        for module in module_payload["modules"]
        for formula_id in module.get("formula_ids", [])
    }
    assert RETIRED_FORMULA_IDS.isdisjoint(module_formula_ids)
    assert "valuation_scenarios" not in {module["module_id"] for module in module_payload["modules"]}
    assert RETIRED_SHEETS.isdisjoint(module_payload["union_sheet_order"])

    shell_manifest = _json(SHELL_MANIFEST)
    assert RETIRED_SHEETS.isdisjoint({row["sheet"] for row in shell_manifest["sheets"]})
    assert all(contract.sheet == "{ticker}_Investment_Case" for contract in USER_INPUT_CONTRACTS)
    assert sum(
        (range_boundaries(contract.target)[2] - range_boundaries(contract.target)[0] + 1)
        * (range_boundaries(contract.target)[3] - range_boundaries(contract.target)[1] + 1)
        for contract in USER_INPUT_CONTRACTS
    ) == 75

    style_text = STYLE_POLICY.read_text(encoding="utf-8")
    assert not any(formula_id in style_text for formula_id in RETIRED_FORMULA_IDS)
    schema = _json(NORMALIZED_SCHEMA)
    fixture = _json(MINIMAL_PACKAGE)
    assert "valuation_outputs" not in schema["properties"]
    assert "valuation_outputs" not in fixture


def test_investment_case_is_the_only_forward_valuation_formula_owner() -> None:
    workbook = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        support = workbook["{ticker}_Investment_Case_Data"]
        assert len(canonical_investment_case_defined_names()) == 40
        for name, (sheet_name, coordinate) in canonical_investment_case_defined_names().items():
            assert workbook.defined_names[name].attr_text.endswith(
                f"'{sheet_name}'!${re.sub(r'[0-9]', '', coordinate)}${re.sub(r'[^0-9]', '', coordinate)}"
            )

        expected_rows = len(CANONICAL_SCENARIOS) * len(CANONICAL_VALUATION_METHODS)
        assert expected_rows == 24
        assert sum(
            cell.data_type == "f"
            for row in support.iter_rows(min_row=2, max_row=25, min_col=57, max_col=67)
            for cell in row
        ) == 264
        for scenario_index in range(4):
            blended_row = 7 + scenario_index * 6
            blended_formula = str(support[f"BI{blended_row}"].value)
            assert "SUM(BN" in blended_formula
            assert "Current share price" not in blended_formula
            assert "$C$106" not in blended_formula
    finally:
        workbook.close()


def test_standard_new_ticker_entrypoints_have_no_legacy_writer_import_edge() -> None:
    entrypoints = (
        ROOT / "pbi_xbrl" / "new_engine_orchestration.py",
        ROOT / "pbi_xbrl" / "new_ticker_value_filler.py",
        ROOT / "pbi_xbrl" / "standard_template_formula_contract.py",
        ROOT / "scripts" / "materialize_standard_template_shell.py",
    )
    for path in entrypoints:
        tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
        imported = {
            alias.name
            for node in ast.walk(tree)
            if isinstance(node, ast.Import)
            for alias in node.names
        }
        imported.update(
            str(node.module)
            for node in ast.walk(tree)
            if isinstance(node, ast.ImportFrom) and node.module
        )
        assert imported.isdisjoint(LEGACY_WRITER_MODULES), (path, imported & LEGACY_WRITER_MODULES)
