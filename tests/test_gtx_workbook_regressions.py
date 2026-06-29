from __future__ import annotations

import datetime as dt
import os
import re
from collections import Counter
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Iterator

import openpyxl
import pytest
from pbi_xbrl.workbook_validation_runner import validate_workbook


DEFAULT_GTX_STAGED_WORKBOOK = Path(
    r"C:\Users\Jibbe\Aktier\StockModelData\outputs\Excel stock models\GTX_model.xlsx"
)
GTX_SPECIAL_EVENT_MARKERS = (
    "eh260781731_ex9902",
    "May 18, 2026",
    "repayment/repricing",
    "debt repayment",
)
GPRE_ONLY_TERMS = (
    "45Z",
    "RVO",
    "E15",
    "Crush margin",
    "crush spread",
    "ethanol",
    "RIN",
)
GTX_CROSS_COMPANY_LEAKAGE_TERMS = (
    "PBI",
    "GPRE",
    "ANF",
    "45Z",
    "RVO",
    "E15",
    "RIN",
    "USPS",
    "Abercrombie",
    "Hollister",
    "Carrier, logistics and service execution reliability",
)
GTX_VISIBLE_UI_SHEETS = (
    "SUMMARY",
    "Valuation",
    "Operating_Drivers",
    "Promise_Progress_UI",
    "Quarter_Notes_UI",
    "GTX_Investment_Case",
    "BS_Segments",
)
GTX_FORBIDDEN_UI_FRAGMENTS = (
    "forward-looking statements",
    "in many cases, you can identify",
    "aim, anticipate, believe, continue, could",
    "variable consideration",
    "revenue is measured as the amount of consideration",
    "pension contribution boilerplate",
    "make contributions of cash and/or marketable securities",
    "each grantee is granted",
    "guidance signal in filing text",
    "margin signal in filing text",
    "revenue signal in filing text",
    "69 >5% average annual cost",
    "cost discipline and flexibility for margins and cash generation 69",
)


def _gtx_workbook_path() -> Path:
    return Path(os.environ.get("GTX_STAGED_WORKBOOK", DEFAULT_GTX_STAGED_WORKBOOK))


@contextmanager
def _gtx_style_workbook() -> Iterator[openpyxl.Workbook]:
    path = _gtx_workbook_path()
    if not path.exists():
        pytest.skip(f"GTX staged workbook not found: {path}")
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    try:
        yield wb
    finally:
        wb.close()


@pytest.fixture()
def gtx_wb() -> openpyxl.Workbook:
    path = _gtx_workbook_path()
    if not path.exists():
        pytest.skip(f"GTX staged workbook not found: {path}")
    return openpyxl.load_workbook(path, data_only=False, read_only=True)


def _sheet_text(wb: openpyxl.Workbook, sheet_name: str) -> str:
    ws = wb[sheet_name]
    return "\n".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )


def _sheet_visible_text(wb: openpyxl.Workbook, sheet_name: str) -> str:
    ws = wb[sheet_name]
    return "\n".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None and not str(cell.value).startswith("=")
    )


def _visible_non_empty_rows(wb: openpyxl.Workbook, sheet_name: str) -> list[list[Any]]:
    ws = wb[sheet_name]
    out: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        values = [value for value in row if value is not None and str(value).strip()]
        if values:
            out.append(values)
    return out


def _used_max_col(wb: openpyxl.Workbook, sheet_name: str) -> int:
    ws = wb[sheet_name]
    max_col = 0
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            if value is not None and str(value).strip():
                max_col = max(max_col, idx)
    return max_col


def _row_texts(wb: openpyxl.Workbook, sheet_name: str) -> list[str]:
    ws = wb[sheet_name]
    out: list[str] = []
    for row in ws.iter_rows(values_only=True):
        values = [str(value) for value in row if value is not None and str(value).strip()]
        if values:
            out.append(" | ".join(values))
    return out


def _history_row(wb: openpyxl.Workbook, quarter: dt.date) -> tuple[list[Any], tuple[Any, ...]]:
    ws = wb["History_Q"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    for row in rows[1:]:
        row_q = row[headers.index("quarter")] if "quarter" in headers else None
        if getattr(row_q, "date", lambda: row_q)() == quarter:
            return headers, row
        if str(row_q).startswith(quarter.isoformat()):
            return headers, row
    raise AssertionError(f"History_Q missing quarter {quarter.isoformat()}")


def _history_value(wb: openpyxl.Workbook, quarter: dt.date, column: str) -> Any:
    headers, row = _history_row(wb, quarter)
    assert column in headers, f"History_Q missing column {column}"
    return row[headers.index(column)]


def _find_row_with_first_cell(ws: Any, label: str) -> int:
    needle = label.strip().lower()
    for row_idx in range(1, int(ws.max_row or 0) + 1):
        if str(ws.cell(row_idx, 1).value or "").strip().lower() == needle:
            return row_idx
    raise AssertionError(f"{ws.title}: missing row label {label!r}")


def _row_values(ws: Any, row_idx: int, *, max_col: int = 15) -> list[Any]:
    return [ws.cell(row_idx, cc).value for cc in range(1, max_col + 1)]


def _font_size_counts(ws: Any, *, max_rows: int = 400, max_cols: int = 20) -> Counter[float]:
    counts: Counter[float] = Counter()
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(int(ws.max_row or 0), max_rows),
        min_col=1,
        max_col=min(int(ws.max_column or 0), max_cols),
    ):
        for cell in row:
            if cell.value is None or str(cell.value).strip() == "":
                continue
            if cell.font and cell.font.sz:
                counts[float(cell.font.sz)] += 1
    return counts


def _rgb(cell: Any) -> str:
    return str(cell.fill.fgColor.rgb or "").upper()


def _non_empty_row_indices(ws: Any, *, max_col: int = 20) -> list[int]:
    rows: list[int] = []
    for rr in range(1, int(ws.max_row or 0) + 1):
        if any(str(ws.cell(rr, cc).value or "").strip() for cc in range(1, min(int(ws.max_column or 0), max_col) + 1)):
            rows.append(rr)
    return rows


def _quarter_note_titles(ws: Any) -> list[str]:
    return [
        str(ws.cell(rr, 1).value or "").strip()
        for rr in range(1, int(ws.max_row or 0) + 1)
        if str(ws.cell(rr, 1).value or "").strip().endswith("- Quarter Notes")
    ]


def _numeric_row_values(ws: Any, row_idx: int, *, max_col: int = 15) -> list[float]:
    values: list[float] = []
    for value in _row_values(ws, row_idx, max_col=max_col):
        if isinstance(value, (int, float)):
            values.append(float(value))
    return values


def test_gtx_staged_workbook_includes_operating_drivers(gtx_wb: openpyxl.Workbook) -> None:
    assert "Operating_Drivers" in gtx_wb.sheetnames


def test_gtx_operating_drivers_has_user_facing_source_backed_watchlist(
    gtx_wb: openpyxl.Workbook,
) -> None:
    assert "Operating_Drivers" in gtx_wb.sheetnames
    rows = _visible_non_empty_rows(gtx_wb, "Operating_Drivers")
    text = _sheet_visible_text(gtx_wb, "Operating_Drivers")

    assert "No operating-driver history available" not in text
    assert len(rows) >= 10
    for needle in (
        "OEM production",
        "product mix",
        "Commercial vehicle",
        "Aftermarket",
        "China",
        "Europe",
        "customer concentration",
        "Adjusted EBIT",
        "Adjusted FCF",
    ):
        assert needle.lower() in text.lower()


def test_gtx_operating_drivers_uses_broad_pbi_style_sections(
    gtx_wb: openpyxl.Workbook,
) -> None:
    assert "Operating_Drivers" in gtx_wb.sheetnames
    rows = _visible_non_empty_rows(gtx_wb, "Operating_Drivers")
    text = _sheet_visible_text(gtx_wb, "Operating_Drivers")

    assert len(rows) >= 40
    assert _used_max_col(gtx_wb, "Operating_Drivers") >= 12
    for section in (
        "Current watchlist",
        "Current/latest outlook",
        "Recent quarter commentary",
        "Data tables",
        "Product-line revenue history",
        "Geography revenue history",
        "Customer concentration",
        "Debt / buyback / leverage watch",
    ):
        assert section.lower() in text.lower()
    assert all(len(row_text) < 700 for row_text in _row_texts(gtx_wb, "Operating_Drivers"))


def test_gtx_operating_drivers_matches_pbi_anf_top_style_conventions(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]

        assert str(ws["A2"].value or "").strip() == "Operating Drivers"
        assert str(ws["A2"].fill.fgColor.rgb or "").upper() == "006FA8DC"
        assert float(ws["A2"].font.sz or 0) == pytest.approx(15.0)
        assert bool(ws["A2"].font.bold)
        assert str(ws["A4"].value or "").strip() == "Current watchlist"
        assert ws.row_dimensions[2].height == pytest.approx(24.0)
        assert ws.row_dimensions[4].height == pytest.approx(22.5)


def test_gtx_operating_drivers_uses_peer_template_blank_context_row(
    gtx_wb: openpyxl.Workbook,
) -> None:
    """GTX should inject content into the peer template, not add a custom subtitle row."""
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]

        assert str(ws["A2"].value or "").strip() == "Operating Drivers"
        assert str(ws["A3"].value or "").strip() == ""
        assert str(ws["A4"].value or "").strip() == "Current watchlist"

        top_rows = [
            float(ws.row_dimensions[row_idx].height or 0.0)
            for row_idx in (2, 4, 5)
        ]
        assert top_rows == pytest.approx([24.0, 22.5, 21.0])


def test_gtx_operating_drivers_exact_peer_style_rhythm(
    gtx_wb: openpyxl.Workbook,
) -> None:
    """GTX Operating_Drivers must use the peer visual contract, not a near-match."""
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]

        assert _rgb(ws["A2"]) == "006FA8DC"
        assert _rgb(ws["A4"]) == "006FA8DC"
        assert _rgb(ws["A5"]) == "00EAF3FB"
        assert float(ws["A4"].font.sz or 0) == pytest.approx(13.0)
        assert float(ws["A5"].font.sz or 0) == pytest.approx(13.0)
        assert ws.row_dimensions[3].height in (None, 0)

        # Peers use 22.5pt body rhythm and white/F7F9FC zebra fills.
        for rr, expected_fill in (
            (6, "00FFFFFF"),
            (7, "00F7F9FC"),
            (8, "00FFFFFF"),
            (9, "00F7F9FC"),
            (10, "00FFFFFF"),
        ):
            assert float(ws.row_dimensions[rr].height or 0) == pytest.approx(22.5)
            assert _rgb(ws.cell(rr, 1)) == expected_fill
            assert float(ws.cell(rr, 1).font.sz or 0) == pytest.approx(12.0)


def test_gtx_operating_drivers_watchlist_uses_anf_pbi_broad_merges(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}

        assert str(ws["A5"].value or "").strip() == "Watch item"
        assert str(ws["B5"].value or "").strip() == "Current read"
        assert str(ws["H5"].value or "").strip() == "Why it matters"
        for row_idx in range(5, 11):
            assert f"B{row_idx}:G{row_idx}" in merged_ranges
            assert f"H{row_idx}:N{row_idx}" in merged_ranges


def test_gtx_operating_driver_watchlist_cells_are_readable(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Operating_Drivers"]
    long_cells: list[tuple[str, int]] = []
    for row in ws.iter_rows(min_row=4, max_row=13, min_col=1, max_col=5):
        for cell in row:
            value = str(cell.value or "").strip()
            if len(value) > 110:
                long_cells.append((cell.coordinate, len(value)))

    assert long_cells == []


def test_gtx_investment_case_has_no_gpre_only_terms(gtx_wb: openpyxl.Workbook) -> None:
    assert "GTX_Investment_Case" in gtx_wb.sheetnames
    text = _sheet_visible_text(gtx_wb, "GTX_Investment_Case").lower()

    leaked: list[str] = []
    for term in GPRE_ONLY_TERMS:
        if " " in term:
            if term.lower() in text:
                leaked.append(term)
            continue
        if re.search(rf"(?<![A-Za-z0-9]){re.escape(term)}(?![A-Za-z0-9])", text, re.I):
            leaked.append(term)

    assert leaked == []


def test_gtx_investment_case_is_gtx_specific_not_generic(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "GTX_Investment_Case")
    low = text.lower()

    assert "generic fallback only" not in low
    assert "pending ticker-specific configuration" not in low
    for needle in (
        "Model read",
        "Why it can work",
        "Key debate",
        "What would improve",
        "What would break",
        "Watch next",
        "Bear",
        "Base",
        "Bull",
        "Quality of Earnings",
        "Adjusted EBIT",
        "Adjusted FCF",
        "turbo",
        "customer concentration",
    ):
        assert needle.lower() in low


def test_gtx_summary_uses_gtx_specific_revenue_and_dependency_language(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "SUMMARY")
    low = text.lower()

    assert "carrier, logistics and service execution reliability" not in low
    assert "operates through accounting view" not in low
    assert "operates through accounting view and product-line mix" not in low
    assert "business model / revenue streams" in low
    assert "n/a\nkey dependencies" not in low
    for needle in (
        "Gas",
        "Diesel",
        "Commercial Vehicle",
        "Aftermarket",
        "Europe",
        "China",
        "Stellantis",
        "BMW",
        "Ford",
    ):
        assert needle.lower() in low


def test_gtx_summary_as_of_quarter_is_human_label_not_excel_serial(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["SUMMARY"]
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if values and str(values[0]).strip().lower() == "as of quarter":
            rendered = values[1]
            assert not isinstance(rendered, (int, float))
            assert not isinstance(rendered, dt.datetime)
            assert str(rendered).strip() in {"2026-Q1", "2026-03-31"}
            break
    else:
        raise AssertionError("SUMMARY missing As of quarter row")


def test_gtx_investment_case_data_has_meaningful_gtx_sections(
    gtx_wb: openpyxl.Workbook,
) -> None:
    assert "GTX_Investment_Case_Data" in gtx_wb.sheetnames
    rows = _visible_non_empty_rows(gtx_wb, "GTX_Investment_Case_Data")
    text = _sheet_visible_text(gtx_wb, "GTX_Investment_Case_Data")
    low = text.lower()

    assert len(rows) >= 25
    assert "generic fallback only" not in low
    for section in (
        "Investment Snapshot",
        "Key Debates",
        "Bear / Base / Bull Scenario",
        "Quality of Earnings",
        "Operating Driver Watchlist",
        "Product / Geography / Customer Cuts",
        "Current Guide -> Implied Earnings",
    ):
        assert section.lower() in low


def test_gtx_investment_case_data_keeps_source_backed_q1_metrics_visible(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "GTX_Investment_Case_Data")
    low = text.lower()

    assert "adjusted ebit $151.0m" in low or "adjusted ebit $151m" in low
    assert "adjusted ebitda $183.0m" in low or "adjusted ebitda $183m" in low
    assert "adjusted fcf $49.0m" in low or "adjusted fcf $49m" in low
    assert "restricted cash" in low
    assert "$2.0m" in low or "$2m" in low


def test_gtx_visible_text_has_no_cross_company_or_sector_leakage(
    gtx_wb: openpyxl.Workbook,
) -> None:
    visible_sheets = (
        "SUMMARY",
        "Operating_Drivers",
        "GTX_Investment_Case",
        "GTX_Investment_Case_Data",
        "Promise_Progress_UI",
    )
    text = "\n".join(_sheet_visible_text(gtx_wb, sheet) for sheet in visible_sheets if sheet in gtx_wb.sheetnames)
    leaked: list[str] = []
    for term in GTX_CROSS_COMPANY_LEAKAGE_TERMS:
        if " " in term or "," in term:
            if term.lower() in text.lower():
                leaked.append(term)
            continue
        if re.search(rf"(?<![A-Za-z0-9]){re.escape(term)}(?![A-Za-z0-9])", text, re.I):
            leaked.append(term)

    assert leaked == []


def test_gtx_promise_progress_has_management_dashboard_guidance_and_actuals(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "Promise_Progress_UI")
    low = text.lower()

    for section in (
        "Management Credibility Scorecard",
        "2026 open guidance",
        "2025 completed actuals",
        "Quarterly guidance timeline / revision log",
        "Post-quarter May 2026 debt repayment/repricing event",
    ):
        assert section.lower() in low

    for required in (
        "Net sales",
        "$3.6bn-$3.9bn",
        "Adjusted EBIT",
        "$520m-$600m",
        "Adjusted FCF",
        "$355m-$475m",
        "FY2025 net sales",
        "$3.584bn",
        "FY2025 buybacks",
        "$208m",
        "common share count reduction",
        "8%",
    ):
        assert required.lower() in low


def test_gtx_guidance_normalized_includes_recent_official_outlook_rows(
    gtx_wb: openpyxl.Workbook,
) -> None:
    assert "Guidance_Normalized" in gtx_wb.sheetnames
    text = _sheet_visible_text(gtx_wb, "Guidance_Normalized")
    low = text.lower()

    assert "2025-q4" in low
    assert "2026-q1" in low
    for required in (
        "net sales",
        "$3.6bn-$3.9bn",
        "adjusted ebit",
        "$520m-$600m",
        "adjusted fcf",
        "$355m-$475m",
        "fy2025 net sales",
        "$3.584bn",
        "fy2025 adjusted ebit",
        "$510m",
        "fy2025 adjusted fcf",
        "$403m",
    ):
        assert required in low


def test_gtx_valuation_guidance_panel_uses_curated_guidance_rows(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Valuation"]
    text = _sheet_visible_text(gtx_wb, "Valuation")
    low = text.lower()

    assert "guidance (as of 2026-03-31)" in low
    assert "guidance (as of 2025-12-31)" in low
    assert "no guidance items for this quarter" not in low
    guidance_header_row = None
    for rr in range(1, int(ws.max_row or 0) + 1):
        row_values = [str(ws.cell(rr, cc).value or "").strip() for cc in range(1, int(ws.max_column or 0) + 1)]
        if any(value.startswith("Guidance (As of 2026-03-31)") for value in row_values):
            guidance_header_row = rr + 1
            break
    assert guidance_header_row is not None
    guidance_headers = [
        str(ws.cell(guidance_header_row, cc).value or "").strip()
        for cc in range(15, 29)
        if str(ws.cell(guidance_header_row, cc).value or "").strip()
    ]
    assert guidance_headers[:4] == ["Metric", "Stated in", "Applies to", "Guidance"]
    assert guidance_headers.count("Metric") == 1
    assert guidance_headers.count("Guidance") == 1
    for required in (
        "net sales",
        "$3.6bn-$3.9bn",
        "constant-currency sales growth",
        "-2% to +6%",
        "net income",
        "$300m-$360m",
        "adjusted ebit",
        "$520m-$600m",
        "cfo",
        "$407m-$522m",
        "adjusted fcf",
        "$355m-$475m",
        "light vehicle production",
        "down 1%-3%",
        "commercial vehicle industry",
        "up 1%-2%",
        "bev penetration",
        "~19%",
        "eur/usd",
        "1.17",
        "rd&e",
        "4.2% of sales",
        "capex",
        "2.5% of sales",
        "fy2025 actuals",
        "$3.584bn",
    ):
        assert required.lower() in low


def test_gtx_valuation_guidance_panel_uses_peer_typography(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Valuation"]
        panel_cells: list[str] = []
        small_cells: list[tuple[str, Any, float]] = []
        for row_idx in range(7, 36):
            for col_idx in range(15, 29):
                cell = ws.cell(row_idx, col_idx)
                value = str(cell.value or "").strip()
                if not value:
                    continue
                panel_cells.append(cell.coordinate)
                font_size = float(cell.font.sz or 0.0)
                if font_size < 12.0:
                    small_cells.append((cell.coordinate, value[:80], font_size))

        assert panel_cells
        assert small_cells == []


def test_gtx_valuation_guidance_panel_uses_peer_fills(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Valuation"]

        for coord in ("O7", "O22", "O29"):
            assert _rgb(ws[coord]) == "006FA8DC", coord
        for coord in ("O8", "O23", "O30"):
            assert _rgb(ws[coord]) == "00EAF3FB", coord
        assert _rgb(ws["O9"]) == "00F7FAFC"
        assert _rgb(ws["O10"]) == "00FFFFFF"


def test_gtx_valuation_operating_driver_panel_is_populated(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "Valuation")
    low = text.lower()

    assert "operating drivers" in low
    assert "no operating-driver map available for this ticker yet" not in low
    for required in (
        "OEM production / end-market demand",
        "Product mix / turbo demand",
        "Commercial vehicle / industrial",
        "Aftermarket",
        "China / Europe exposure",
        "Customer concentration",
        "RD&E / technology awards",
        "Adjusted EBIT / adjusted FCF conversion",
        "Debt, net leverage and buybacks",
    ):
        assert required.lower() in low


def test_gtx_valuation_capital_return_rows_have_no_year_artifacts(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Valuation"]
    labels = {"Buybacks (TTM, cash)", "Dividends (TTM, cash)"}
    bad: list[tuple[str, str, Any]] = []
    found: set[str] = set()
    for row in ws.iter_rows():
        label = str(row[0].value or "").strip() if row else ""
        if label not in labels:
            continue
        found.add(label)
        for cell in row[1:13]:
            value = cell.value
            if isinstance(value, (int, float)) and float(value).is_integer() and 1900 <= int(value) <= 2100:
                bad.append((label, cell.coordinate, value))

    assert found == labels
    assert bad == []


def test_gtx_valuation_adjusted_metric_ttm_rows_follow_base_metrics(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Valuation"]

    assert _find_row_with_first_cell(ws, "Adj EBIT (TTM)") == _find_row_with_first_cell(ws, "Adj EBIT") + 1
    assert _find_row_with_first_cell(ws, "Adj EBITDA (TTM)") == _find_row_with_first_cell(ws, "Adj EBITDA") + 1
    assert _find_row_with_first_cell(ws, "Adj FCF (TTM)") == _find_row_with_first_cell(ws, "Adj FCF") + 1

    adjusted_sequence = [
        "Adj EBIT",
        "Adj EBIT (TTM)",
        "Adj EBIT margin %",
        "Adj EBITDA",
        "Adj EBITDA (TTM)",
        "Adj EBITDA margin %",
        "Adj FCF",
        "Adj FCF (TTM)",
    ]
    adjusted_rows = [_find_row_with_first_cell(ws, label) for label in adjusted_sequence]
    assert adjusted_rows == list(range(adjusted_rows[0], adjusted_rows[0] + len(adjusted_rows)))


def test_gtx_operating_drivers_recent_commentary_uses_broad_layout(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]

    header_row = None
    for row_idx in range(1, int(ws.max_row or 0) + 1):
        if "Recent quarter commentary" in str(ws.cell(row_idx, 1).value or ""):
            header_row = row_idx + 1
            break
    assert header_row is not None

    headers = [str(ws.cell(header_row, cc).value or "").strip() for cc in range(1, 15)]
    assert headers[:3] == ["Horizon", "Stated in", "Commentary"]
    assert not any("source" in value.lower() for value in headers[:8])
    assert not any("confidence" in value.lower() for value in headers[:8])

    commentary_rows = [
        rr
        for rr in range(header_row + 1, int(ws.max_row or 0) + 1)
        if str(ws.cell(rr, 2).value or "").strip().startswith(("2026-Q1", "2025-Q", "2024-Q", "2023-Q"))
    ]
    assert len(commentary_rows) >= 12
    commentary_text = _sheet_visible_text(gtx_wb, "Operating_Drivers").lower()
    for expected in ("2026-q1", "2025-q4", "2025-q3", "2025-q2", "2025-q1", "2024-q4"):
        assert expected in commentary_text
    assert "official transcripts not loaded" in commentary_text
    assert "earnings release / presentation text only" in commentary_text

    first_commentary_row = next(
        rr
        for rr in range(header_row + 1, int(ws.max_row or 0) + 1)
        if str(ws.cell(rr, 2).value or "").strip()
    )
    assert str(ws.cell(first_commentary_row, 3).value or "").strip()
    assert float(ws.row_dimensions[first_commentary_row].height or 0.0) <= 24.5


def test_gtx_operating_drivers_recent_commentary_text_spans_wide_area(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
        section_row = _find_row_with_first_cell(
            ws,
            "Recent quarter commentary \u2014 source-backed actuals and management framing",
        )
        header_row = section_row + 1
        assert f"C{header_row}:N{header_row}" in merged_ranges

        data_rows = [
            row_idx
            for row_idx in range(header_row + 1, int(ws.max_row or 0) + 1)
            if str(ws.cell(row_idx, 2).value or "").strip().startswith(("2026-Q", "2025-Q", "2024-Q", "2023-Q"))
        ]
        assert data_rows
        for row_idx in data_rows[:12]:
            assert f"C{row_idx}:N{row_idx}" in merged_ranges


def test_gtx_operating_drivers_uses_uniform_peer_widths_and_right_side_audit_columns(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]
        widths = [float(ws.column_dimensions[col].width or 0.0) for col in "BCDEFGHIJKLMN"]
        assert min(widths) >= 13.0
        assert max(widths) - min(widths) <= 1.0

        misplaced: list[tuple[int, int, str]] = []
        audit_terms = {"source", "evidence", "workbook treatment", "confidence", "audit note"}
        for rr in range(1, int(ws.max_row or 0) + 1):
            for cc in range(2, 9):
                value = str(ws.cell(rr, cc).value or "").strip().lower()
                if value in audit_terms:
                    misplaced.append((rr, cc, value))

        assert misplaced == []


def test_gtx_operating_drivers_table_columns_are_wide_enough_for_labels(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]
        widths = {col: float(ws.column_dimensions[col].width or 0.0) for col in ("B", "C", "D", "E", "F", "G", "H", "I")}

    assert min(widths.values()) >= 13.0
    assert max(widths.values()) - min(widths.values()) <= 1.0


def test_gtx_operating_drivers_outlook_uses_peer_topic_read_source_template(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
        section_row = _find_row_with_first_cell(ws, "Current/latest outlook")
        header_row = section_row + 1
        headers = [str(ws.cell(header_row, cc).value or "").strip() for cc in range(1, 15)]

        assert "Unit" not in headers
        assert "Low" not in headers
        assert "High" not in headers
        assert headers[0] == "Topic"
        assert headers[1] == "Current read"
        assert headers[7] == "Source / use"
        assert f"B{header_row}:G{header_row}" in merged_ranges
        assert f"H{header_row}:N{header_row}" in merged_ranges
        first_data_row = header_row + 1
        assert f"B{first_data_row}:G{first_data_row}" in merged_ranges
        assert f"H{first_data_row}:N{first_data_row}" in merged_ranges


def test_gtx_operating_drivers_has_no_freeze_panes_and_release_commentary_note(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]
        text = _sheet_visible_text(full_wb, "Operating_Drivers").lower()

        assert ws.freeze_panes in (None, "")
        assert "management / release commentary" in text
        assert "official transcripts not loaded" in text
        assert "earnings release / presentation text only" in text


def test_gtx_operating_drivers_peer_font_and_row_height_pattern(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]
        font_counts = _font_size_counts(ws, max_rows=160, max_cols=14)
        row_heights = [
            float(ws.row_dimensions[row_idx].height or 15.0)
            for row_idx in range(1, int(ws.max_row or 0) + 1)
        ]

        assert font_counts[12.0] >= 250
        assert font_counts[13.0] >= 20
        assert font_counts[10.5] <= 5
        assert max(row_heights) <= 24.5
        assert 18.0 <= (sum(row_heights) / len(row_heights)) <= 23.0


def test_gtx_operating_drivers_outlook_values_inline_units_and_clean_wording(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Operating_Drivers"]
    section_row = _find_row_with_first_cell(ws, "Current/latest outlook")
    visible_text = "\n".join(
        " ".join(str(ws.cell(rr, cc).value or "").strip() for cc in range(1, 15))
        for rr in range(section_row, section_row + 10)
    ).lower()

    for required in ("$3.6bn", "$3.9bn", "-2%", "+6%", "~19%", "4.2% of sales", "2.5% of sales"):
        assert required in visible_text
    assert "about 19%" not in visible_text
    assert "about 4.2%" not in visible_text
    assert "about 2.5%" not in visible_text


def test_gtx_operating_drivers_debt_watch_uses_broad_business_columns(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
        section_row = _find_row_with_first_cell(ws, "Debt / buyback / leverage watch")
        header_row = section_row + 1
        headers = [str(ws.cell(header_row, cc).value or "").strip() for cc in range(1, 15)]

        assert headers[1] == "Reported / disclosed value"
        assert headers[4] == "Period / event"
        assert headers[6] == "Why it matters"
        assert headers[9] == "Source"
        assert headers[11] == "Workbook treatment / note"
        for row_idx in (header_row, header_row + 1):
            assert f"B{row_idx}:D{row_idx}" in merged_ranges
            assert f"E{row_idx}:F{row_idx}" in merged_ranges
            assert f"G{row_idx}:I{row_idx}" in merged_ranges
            assert f"J{row_idx}:K{row_idx}" in merged_ranges
            assert f"L{row_idx}:N{row_idx}" in merged_ranges


def test_gtx_operating_driver_audit_fields_are_visually_secondary(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]
        checked_cells: list[str] = []
        for section, audit_start_col in (
            ("Current/latest outlook", 8),
            ("Recent quarter commentary \u2014 source-backed actuals and management framing", 9),
            (
                "Management / release commentary \u2014 Official transcripts not loaded; commentary uses earnings release / presentation text only.",
                9,
            ),
            ("Debt / buyback / leverage watch", 10),
        ):
            section_row = _find_row_with_first_cell(ws, section)
            row_idx = section_row + 2
            while row_idx <= int(ws.max_row or 0):
                if row_idx > section_row + 1:
                    first_col = str(ws.cell(row_idx, 1).value or "").strip()
                    first_fill = str(ws.cell(row_idx, 1).fill.fgColor.rgb or "").upper()
                    if first_col and first_fill.endswith("6FA8DC"):
                        break
                if not any(str(ws.cell(row_idx, cc).value or "").strip() for cc in range(1, 15)):
                    break
                for cc in range(audit_start_col, 15):
                    cell = ws.cell(row_idx, cc)
                    if not str(cell.value or "").strip():
                        continue
                    checked_cells.append(cell.coordinate)
                    assert float(cell.font.sz or 0.0) <= 12.0
                    assert str(cell.font.color.rgb or "").upper().endswith("5F6B76")
                row_idx += 1

        assert checked_cells


def test_gtx_operating_driver_tables_are_chronological_older_to_newer(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Operating_Drivers"]

    def _headers_after(section: str) -> list[str]:
        for row_idx in range(1, int(ws.max_row or 0) + 1):
            if str(ws.cell(row_idx, 1).value or "").strip() == section:
                return [str(ws.cell(row_idx + 1, cc).value or "").strip() for cc in range(1, 8)]
        raise AssertionError(f"Missing section {section}")

    assert _headers_after("Product-line revenue history")[:6] == [
        "Product line",
        "2023 year",
        "2024 year",
        "2025 year",
        "2025-Q1",
        "2026-Q1",
    ]
    assert _headers_after("Geography revenue history")[:6] == [
        "Geography",
        "2023 year",
        "2024 year",
        "2025 year",
        "2025-Q1",
        "2026-Q1",
    ]


def test_gtx_operating_driver_data_table_source_treatment_cells_are_wide_merges(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
        for section in (
            "Product-line revenue history",
            "Geography revenue history",
            "Customer concentration",
        ):
            section_row = _find_row_with_first_cell(ws, section)
            header_row = section_row + 1
            assert str(ws.cell(header_row, 10).value or "").strip() == "Source"
            assert str(ws.cell(header_row, 12).value or "").strip() == "Treatment"
            assert f"J{header_row}:K{header_row}" in merged_ranges
            assert f"L{header_row}:N{header_row}" in merged_ranges
            body_rows = [
                rr
                for rr in range(header_row + 1, min(header_row + 8, int(ws.max_row or 0)) + 1)
                if str(ws.cell(rr, 1).value or "").strip()
                and not str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper().endswith("6FA8DC")
            ]
            assert body_rows
            for rr in body_rows:
                assert f"J{rr}:K{rr}" in merged_ranges
                assert f"L{rr}:N{rr}" in merged_ranges


def test_gtx_operating_driver_debt_watch_source_and_treatment_are_wide_merges(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Operating_Drivers"]
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
        section_row = _find_row_with_first_cell(ws, "Debt / buyback / leverage watch")
        header_row = section_row + 1
        assert f"J{header_row}:K{header_row}" in merged_ranges
        assert f"L{header_row}:N{header_row}" in merged_ranges
        for row_idx in range(header_row + 1, header_row + 8):
            assert f"J{row_idx}:K{row_idx}" in merged_ranges
            assert f"L{row_idx}:N{row_idx}" in merged_ranges


def test_gtx_operating_driver_tables_inline_units_in_row_labels(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Operating_Drivers"]

    for section, expected_labels in {
        "Product-line revenue history": ("Gas ($m)", "Diesel ($m)", "Aftermarket ($m)"),
        "Geography revenue history": ("United States ($m)", "Europe ($m)", "China ($m)"),
    }.items():
        section_row = _find_row_with_first_cell(ws, section)
        headers = [str(ws.cell(section_row + 1, cc).value or "").strip() for cc in range(1, 15)]
        assert "Unit" not in headers
        block_text = "\n".join(
            str(ws.cell(rr, 1).value or "")
            for rr in range(section_row + 2, min(section_row + 12, int(ws.max_row or 0)) + 1)
        )
        for label in expected_labels:
            assert label in block_text


def test_gtx_customer_concentration_separates_revenue_and_sales_percent_blocks(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Operating_Drivers"]
    section_row = _find_row_with_first_cell(ws, "Customer concentration")
    labels = [
        str(ws.cell(rr, 1).value or "").strip()
        for rr in range(section_row + 2, min(section_row + 14, int(ws.max_row or 0)) + 1)
    ]

    revenue_positions = {
        label: idx
        for idx, label in enumerate(labels)
        if label.endswith("revenue ($m)") or label == "Top ten customers revenue ($m)"
    }
    percent_positions = {
        label: idx
        for idx, label in enumerate(labels)
        if label.endswith("% sales") or label == "Top ten customers % sales"
    }
    blank_positions = [idx for idx, label in enumerate(labels) if not label]

    assert {"Stellantis revenue ($m)", "BMW revenue ($m)", "Ford revenue ($m)"}.issubset(revenue_positions)
    assert {"Stellantis % sales", "BMW % sales", "Ford % sales", "Top ten customers % sales"}.issubset(percent_positions)
    assert blank_positions, "Expected visual separator between revenue rows and % sales rows"
    assert max(revenue_positions.values()) < min(blank_positions) < min(percent_positions.values())


def test_gtx_bs_segments_analytical_cuts_use_readable_units_and_customer_blocks(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["BS_Segments"]

    product_row = _find_row_with_first_cell(ws, "Product-line revenue by year")
    product_headers = [str(ws.cell(product_row + 1, cc).value or "").strip() for cc in range(1, 10)]
    assert "Unit" not in product_headers
    product_labels = [
        str(ws.cell(rr, 1).value or "").strip()
        for rr in range(product_row + 2, min(product_row + 8, int(ws.max_row or 0)) + 1)
    ]
    assert {"Gas ($m)", "Diesel ($m)", "Aftermarket ($m)"}.issubset(set(product_labels))

    geography_row = _find_row_with_first_cell(ws, "Geography revenue by year")
    geography_headers = [str(ws.cell(geography_row + 1, cc).value or "").strip() for cc in range(1, 10)]
    assert "Unit" not in geography_headers
    geography_labels = [
        str(ws.cell(rr, 1).value or "").strip()
        for rr in range(geography_row + 2, min(geography_row + 8, int(ws.max_row or 0)) + 1)
    ]
    assert {"United States ($m)", "Europe ($m)", "China ($m)"}.issubset(set(geography_labels))

    customer_row = _find_row_with_first_cell(ws, "Customer concentration")
    labels = [
        str(ws.cell(rr, 1).value or "").strip()
        for rr in range(customer_row + 2, min(customer_row + 14, int(ws.max_row or 0)) + 1)
    ]
    revenue_positions = {
        label: idx
        for idx, label in enumerate(labels)
        if label.endswith("revenue ($m)") or label == "Top ten customers revenue ($m)"
    }
    percent_positions = {
        label: idx
        for idx, label in enumerate(labels)
        if label.endswith("% sales") or label == "Top ten customers % sales"
    }
    blank_positions = [idx for idx, label in enumerate(labels) if not label]

    assert {"Stellantis revenue ($m)", "BMW revenue ($m)", "Ford revenue ($m)"}.issubset(revenue_positions)
    assert {"Stellantis % sales", "BMW % sales", "Ford % sales", "Top ten customers % sales"}.issubset(percent_positions)
    assert blank_positions, "Expected visual separator between customer revenue and % sales rows"
    assert max(revenue_positions.values()) < min(blank_positions) < min(percent_positions.values())


def test_gtx_operating_drivers_keeps_adjusted_metric_history_out_of_driver_sheet(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "Operating_Drivers")
    assert "Adjusted EBIT / EBITDA / adjusted FCF history" not in text


def test_gtx_valuation_quarterly_grid_includes_adjusted_operating_and_fcf_rows(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Valuation"]
    text = _sheet_visible_text(gtx_wb, "Valuation")
    low = text.lower()

    for label in ("Adj EBIT", "Adj EBITDA", "Adj FCF"):
        assert label.lower() in low

    adj_ebit_row = _find_row_with_first_cell(ws, "Adj EBIT")
    adj_ebitda_row = _find_row_with_first_cell(ws, "Adj EBITDA")
    adj_fcf_row = _find_row_with_first_cell(ws, "Adj FCF")
    assert any(value == pytest.approx(151.0, abs=0.01) for value in _numeric_row_values(ws, adj_ebit_row))
    assert any(value == pytest.approx(183.0, abs=0.01) for value in _numeric_row_values(ws, adj_ebitda_row))
    assert any(value == pytest.approx(49.0, abs=0.01) for value in _numeric_row_values(ws, adj_fcf_row))


def test_gtx_promise_progress_body_font_matches_ui_convention(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Promise_Progress_UI"]
    mismatches: list[tuple[str, Any]] = []
    for row in ws.iter_rows(min_row=6, max_row=min(int(ws.max_row or 0), 42), min_col=1, max_col=14):
        fg_color = getattr(getattr(row[0], "fill", None), "fgColor", None)
        section_fill = str(getattr(fg_color, "rgb", "") or "").upper()
        if section_fill.endswith(("5B9BD5", "6FA8DC", "4472C4")):
            continue
        for col_idx, cell in enumerate(row, start=1):
            if col_idx in {11, 12}:
                continue
            if cell.value is None or str(cell.value).strip() == "":
                continue
            if float(cell.font.sz or 0) < 11.0:
                mismatches.append((cell.coordinate, cell.font.sz))

    assert mismatches == []


def test_gtx_promise_progress_uses_peer_freeze_and_wide_note_columns(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Promise_Progress_UI"]

        assert ws.freeze_panes in (None, "", "A2")
        assert int(ws.max_column or 0) >= 15
        expected_widths = {
            "A": 28.0,
            "B": 28.0,
            "C": 32.0,
            "D": 15.0,
            "E": 22.0,
            "F": 28.0,
            "G": 15.0,
            "H": 14.0,
            "I": 16.0,
            "J": 14.0,
            "K": 42.0,
            "L": 42.0,
            "M": 4.0,
            "N": 4.0,
        }
        for col, expected in expected_widths.items():
            assert float(ws.column_dimensions[col].width or 0.0) == pytest.approx(expected, abs=0.03)
        o_width = float(ws.column_dimensions["O"].width or 0.0)
        assert o_width == pytest.approx(4.0, abs=0.03) or o_width == pytest.approx(24.0, abs=0.03)


def test_gtx_promise_progress_has_no_validation_quarter_label_issues(
    gtx_wb: openpyxl.Workbook,
) -> None:
    result = validate_workbook(_gtx_workbook_path(), "GTX")
    promise_issues = [issue.detail for issue in result.issues if "Promise_Progress_UI!" in issue.detail]

    assert result.quarter_label_issue_count == 0
    assert promise_issues == []


def test_gtx_promise_progress_uses_anf_pbi_card_layout(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Promise_Progress_UI"]
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}

        assert str(ws["A1"].value or "").strip() == "Promise Progress"
        assert str(ws["A3"].value or "").strip() == "Management Credibility Scorecard"
        assert str(ws["A3"].fill.fgColor.rgb or "").upper() == "005B9BD5"
        assert int(ws.max_row or 0) >= 70
        assert float(ws.row_dimensions[1].height or 0.0) <= 24.0
        assert float(ws.row_dimensions[3].height or 0.0) <= 24.0
        tall_rows = [
            row_idx
            for row_idx in range(1, int(ws.max_row or 0) + 1)
            if float(ws.row_dimensions[row_idx].height or 0.0) > 24.5
        ]
        assert tall_rows == []
        for required_merge in ("A3:L3", "C5:F5", "G5:L5"):
            assert required_merge in merged_ranges


def test_gtx_promise_progress_open_guidance_uses_compact_peer_columns(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Promise_Progress_UI"]
        section_row = _find_row_with_first_cell(ws, "2026 open guidance")
        headers = [
            str(ws.cell(section_row + 1, cc).value or "").strip()
            for cc in range(1, 13)
            if str(ws.cell(section_row + 1, cc).value or "").strip()
        ]

        assert "Current guide" in headers
        assert "Status" in headers
        assert "Notes/source" in headers
        assert "Unit" not in headers
        assert "Source date" not in headers
        assert "Workbook treatment" not in headers
        assert "Confidence" not in headers

        open_rows = [
            rr
            for rr in range(section_row + 2, min(section_row + 16, int(ws.max_row or 0)) + 1)
            if str(ws.cell(rr, 1).value or "").strip()
        ]
        for rr in open_rows:
            last_visible_col = max(
                cc
                for cc in range(1, 15)
                if str(ws.cell(rr, cc).value or "").strip()
            )
            assert last_visible_col <= 12


def test_gtx_promise_progress_status_cells_are_colored_like_peer_dashboards(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Promise_Progress_UI"]
        status_palette = {"0099CCFF", "005B9BD5", "0064C2A6", "00A9D18E", "00F4B183", "00ED7D31"}
        sections = ("2026 open guidance", "2025 completed actuals", "2026-Q1 revisions")

        for section in sections:
            section_row = _find_row_with_first_cell(ws, section)
            headers = {
                str(ws.cell(section_row + 1, cc).value or "").strip(): cc
                for cc in range(1, 13)
            }
            status_col = headers.get("Status")
            assert status_col is not None
            checked = 0
            for rr in range(section_row + 2, int(ws.max_row or 0) + 1):
                if rr > section_row + 2 and str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper().endswith("5B9BD5"):
                    break
                marker = str(ws.cell(rr, 1).value or "").strip()
                if marker in {
                    "2025 completed actuals",
                    "Quarterly actual/context rows",
                    "Quarterly guidance timeline / revision log",
                    "Post-quarter May 2026 debt repayment/repricing event",
                }:
                    break
                if not marker:
                    continue
                status = str(ws.cell(rr, status_col).value or "").strip()
                if not status:
                    continue
                fill = str(ws.cell(rr, status_col).fill.fgColor.rgb or "").upper()
                assert fill in status_palette, (section, rr, status, fill)
                checked += 1
            assert checked >= 1


def test_gtx_promise_progress_core_columns_are_not_cramped_or_audit_dominated(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Promise_Progress_UI"]
        merged_ranges = list(ws.merged_cells.ranges)

        def _merged_width(row_idx: int, col_idx: int) -> int:
            for rng in merged_ranges:
                if rng.min_row <= row_idx <= rng.max_row and rng.min_col <= col_idx <= rng.max_col:
                    return int(rng.max_col - rng.min_col + 1)
            return 1

        audit_terms = ("source-backed", "workbook treatment", "confidence", "8-k package")
        cramped: list[tuple[str, str, int]] = []
        audit_in_core: list[tuple[str, str]] = []
        for rr in range(1, min(int(ws.max_row or 0), 90) + 1):
            for cc in range(1, 8):
                value = str(ws.cell(rr, cc).value or "").strip()
                if not value:
                    continue
                low = value.lower()
                if any(term in low for term in audit_terms):
                    audit_in_core.append((ws.cell(rr, cc).coordinate, value[:90]))
                if len(value) > 95 and _merged_width(rr, cc) < 2:
                    cramped.append((ws.cell(rr, cc).coordinate, value[:90], len(value)))

        assert audit_in_core == []
        assert cramped == []


def test_gtx_promise_progress_source_notes_are_right_side_and_secondary(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Promise_Progress_UI"]
        checked_cells: list[str] = []
        for rr in range(1, min(int(ws.max_row or 0), 90) + 1):
            for cc in (11, 12):
                cell = ws.cell(rr, cc)
                value = str(cell.value or "").strip()
                if not value:
                    continue
                if str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper().endswith(("5B9BD5", "6FA8DC", "4472C4")):
                    continue
                if bool(cell.font.bold):
                    continue
                checked_cells.append(cell.coordinate)
                assert float(cell.font.sz or 0.0) <= 11.0
                assert str(cell.font.color.rgb or "").upper().endswith("5F6B76")

        assert checked_cells


def test_gtx_promise_progress_has_quarter_context_beyond_two_topline_sections(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "Promise_Progress_UI").lower()

    for required in (
        "2025 guidance progression",
        "2024 guidance progression",
        "quarterly guidance timeline / revision log",
        "2026-q1",
        "2025-q4",
        "2025-q3",
        "2025-q2",
        "2025-q1",
        "only clean official guidance revisions are shown",
    ):
        assert required in text


def test_gtx_promise_progress_timeline_uses_peer_quarter_revision_blocks(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Promise_Progress_UI"]
        timeline_row = _find_row_with_first_cell(ws, "Quarterly guidance timeline / revision log")
        required_blocks = [
            "2026-Q1 revisions",
            "2025-Q4 revisions",
            "2025-Q3 revisions",
            "2025-Q2 revisions",
            "2025-Q1 revisions",
            "2024-Q4 revisions",
            "2024-Q3 revisions",
            "2024-Q2 revisions",
            "2024-Q1 revisions",
            "2023-Q4 revisions",
            "2023-Q3 revisions",
            "2023-Q2 revisions",
            "2023-Q1 revisions",
            "2022-Q4 revisions",
        ]
        block_rows: dict[str, int] = {}
        for row_idx in range(timeline_row + 1, int(ws.max_row or 0) + 1):
            value = str(ws.cell(row_idx, 1).value or "").strip()
            if value in required_blocks:
                block_rows[value] = row_idx

        assert list(block_rows) == required_blocks
        for block, row_idx in block_rows.items():
            assert str(ws.cell(row_idx, 1).fill.fgColor.rgb or "").upper().endswith("5B9BD5"), block
            next_row = row_idx + 1
            headers = [
                str(ws.cell(next_row, cc).value or "").strip()
                for cc in range(1, 13)
                if str(ws.cell(next_row, cc).value or "").strip()
            ]
            assert headers[:3] == ["Metric", "Previous guide", "New/current guide"], block


def test_gtx_promise_progress_source_backed_rows_keep_hidden_source_keys(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Promise_Progress_UI"]
        missing: list[str] = []

        for rr in range(1, int(ws.max_row or 0) + 1):
            first = str(ws.cell(rr, 1).value or "").strip()
            if not first or first in {"Metric", "Event"}:
                continue
            if not (first.endswith("guidance") or first.endswith("actuals") or first == "Debt reduction"):
                continue
            source_blob = " ".join(str(ws.cell(rr, cc).value or "") for cc in range(10, 13)).lower()
            if not any(token in source_blob for token in ("release", "history_q", "source-backed", "8-k", "10-q")):
                continue
            hidden_key = str(ws.cell(rr, 15).value or "").strip()
            if not hidden_key.startswith("guidance:"):
                missing.append(f"{ws.title}!O{rr} for {first!r}: {hidden_key!r}")

        assert missing == []


def test_gtx_promise_progress_density_and_row_heights_match_peer_dashboard(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Promise_Progress_UI"]
        non_empty = _non_empty_row_indices(ws, max_col=15)
        font_counts = _font_size_counts(ws, max_rows=140, max_cols=15)
        row_heights = [
            float(ws.row_dimensions[row_idx].height or 15.0)
            for row_idx in range(1, int(ws.max_row or 0) + 1)
        ]

        assert len(non_empty) >= 65
        assert font_counts[11.0] >= 350
        assert font_counts[13.0] + font_counts[14.0] <= 5
        assert max(row_heights) <= 24.5
        assert sum(1 for height in row_heights if height > 26.0) == 0


def test_gtx_promise_progress_top_spacing_matches_peer_dashboard_template(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Promise_Progress_UI"]

        assert str(ws["A1"].value or "").strip() == "Promise Progress"
        assert str(ws["A3"].value or "").strip() == "Management Credibility Scorecard"
        assert str(ws["A11"].value or "").strip().endswith("guidance progression")
        assert str(ws["A21"].value or "").strip() == "2026 open guidance"
        assert float(ws.column_dimensions["O"].width or 0.0) >= 13.0
        assert [
            float(ws.row_dimensions[row_idx].height or 0.0)
            for row_idx in range(1, 13)
        ] == pytest.approx([24.0, 24.0, 22.0, 22.0, 24.0, 24.0, 24.0, 24.0, 24.0, 18.0, 22.0, 22.0])


def test_gtx_promise_progress_uses_exact_peer_fills_and_body_rhythm(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Promise_Progress_UI"]

        assert _rgb(ws["A1"]) == "005B9BD5"
        assert _rgb(ws["A2"]) == "00F6F9FC"
        assert _rgb(ws["A3"]) == "005B9BD5"
        assert _rgb(ws["A4"]) == "00EAF3FB"
        assert _rgb(ws["A5"]) == "00F6F9FC"
        assert _rgb(ws["A6"]) == "00FFFFFF"
        assert _rgb(ws["A7"]) == "00F6F9FC"

        # Body rows in the dashboard use 24pt rhythm; section/header rows use 22pt.
        assert float(ws.row_dimensions[13].height or 0) == pytest.approx(24.0)
        assert float(ws.row_dimensions[14].height or 0) == pytest.approx(24.0)
        assert float(ws.row_dimensions[23].height or 0) == pytest.approx(24.0)
        assert float(ws["A1"].font.sz or 0) == pytest.approx(12.0)
        assert float(ws["A4"].font.sz or 0) == pytest.approx(11.0)


def test_gtx_bs_segments_shows_analytical_cuts_not_missing_segment_stub(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "BS_Segments")
    low = text.lower()

    assert "no annual segment data found for gtx" not in low
    assert "annual analytical cuts" in low
    assert low.count("annual analytical cuts") == 1
    assert "one operating/reportable segment" in low
    assert "not segment profit" in low
    for required in (
        "Gas",
        "Diesel",
        "Commercial Vehicles / Industrial",
        "Aftermarket",
        "Other",
        "United States",
        "Europe",
        "China",
        "Rest of Asia",
        "Other International",
        "Stellantis",
        "BMW",
        "Ford",
    ):
        assert required.lower() in low


def test_gtx_bs_segments_analytical_cut_columns_are_not_overwide(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["BS_Segments"]
        assert float(ws.column_dimensions["A"].width or 0.0) == pytest.approx(54.0, abs=0.03)
        widths = [float(ws.column_dimensions[col].width or 0.0) for col in ("B", "C", "D", "E", "F", "G", "H", "I")]

        assert all(width == pytest.approx(11.29, abs=0.03) for width in widths)
        body_sizes = {
            float(ws.cell(rr, cc).font.sz or 0.0)
            for rr in range(59, min(86, int(ws.max_row or 0)) + 1)
            for cc in range(1, 10)
            if ws.cell(rr, cc).value not in (None, "")
        }
        assert any(size >= 12.0 for size in body_sizes)


def test_gtx_bs_segments_source_treatment_uses_merged_note_area(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["BS_Segments"]
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
        product_row = _find_row_with_first_cell(ws, "Product-line revenue by year")
        header_row = product_row + 1
        first_data_row = product_row + 2

        headers = [str(ws.cell(header_row, cc).value or "").strip() for cc in range(1, 10)]
        assert headers[0] == "Product line"
        assert headers[1:4] == ["2023 year", "2024 year", "2025 year"]
        assert headers[4] == "Source / treatment"
        assert "Source" not in headers[5:]
        assert "Treatment" not in headers[5:]
        assert f"E{header_row}:I{header_row}" in merged_ranges
        assert f"E{first_data_row}:I{first_data_row}" in merged_ranges
        assert "2025 Form 10-K" in str(ws.cell(first_data_row, 5).value or "")


def test_gtx_investment_case_key_columns_have_room_for_readability(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["GTX_Investment_Case"]
        widths = {col: float(ws.column_dimensions[col].width or 0.0) for col in "ABCDEFGHIJKLMN"}

    assert widths["A"] + widths["B"] >= 54.0
    assert widths["B"] + widths["C"] + widths["D"] >= 70.0
    assert sum(widths[col] for col in "HIJKLMN") >= 120.0


def test_gtx_investment_case_key_sections_use_broad_reading_ranges(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["GTX_Investment_Case"]
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}

        bridge_header = _find_row_with_first_cell(ws, "Bridge item")
        assert f"H{bridge_header}:N{bridge_header}" in merged_ranges

        quality_row = _find_row_with_first_cell(ws, "Quality of Earnings")
        quality_header = quality_row + 1
        assert f"B{quality_header}:D{quality_header}" in merged_ranges
        assert f"G{quality_header}:N{quality_header}" in merged_ranges

        work_row = _find_row_with_first_cell(ws, "What needs to happen for the stock to work")
        work_header = work_row + 1
        assert f"A{work_header}:B{work_header}" in merged_ranges
        assert f"C{work_header}:J{work_header}" in merged_ranges


def test_gtx_valuation_has_no_guidance_or_driver_fallback_text(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "Valuation").lower()

    assert "no guidance items for this quarter" not in text
    assert "no operating-driver map available for this ticker yet" not in text


def test_gtx_visible_ui_sheets_filter_boilerplate_ocr_and_policy_noise(
    gtx_wb: openpyxl.Workbook,
) -> None:
    found: list[tuple[str, str]] = []
    for sheet_name in GTX_VISIBLE_UI_SHEETS:
        assert sheet_name in gtx_wb.sheetnames
        text = _sheet_visible_text(gtx_wb, sheet_name).lower()
        for needle in GTX_FORBIDDEN_UI_FRAGMENTS:
            if needle.lower() in text:
                found.append((sheet_name, needle))

    assert found == []


def test_gtx_quarter_narrative_data_does_not_mark_noisy_rows_ui_eligible(
    gtx_wb: openpyxl.Workbook,
) -> None:
    assert "Quarter_Narrative_Data" in gtx_wb.sheetnames
    ws = gtx_wb["Quarter_Narrative_Data"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    include_idx = headers.index("Include in UI")

    bad_rows: list[tuple[int, str]] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        include_value = str(row[include_idx] or "").strip().lower()
        if include_value not in {"yes", "true", "1"}:
            continue
        row_text = " ".join(str(value) for value in row if value not in (None, "")).lower()
        for needle in GTX_FORBIDDEN_UI_FRAGMENTS:
            if needle.lower() in row_text:
                bad_rows.append((row_idx, needle))

    assert bad_rows == []


def test_gtx_quarter_notes_filters_boilerplate_ocr_and_policy_noise(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "Quarter_Notes_UI").lower()
    forbidden = (
        "forward-looking statements",
        "in many cases, you can identify",
        "aim,",
        "anticipate",
        "believe,",
        "continue,",
        "could",
        "variable consideration",
        "revenue is measured as the amount of consideration",
        "pension contribution",
        "make contributions of cash and/or marketable securities",
        "each grantee is granted",
        "guidance signal in filing text",
        "margin signal in filing text",
        "revenue signal in filing text",
        "69 >5% average annual cost",
    )
    found = [needle for needle in forbidden if needle in text]

    assert found == []


def test_gtx_quarter_notes_uses_clean_source_backed_latest_quarter_reads(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "Quarter_Notes_UI").lower()

    for required in (
        "2026-q1 - quarter notes",
        "net sales $985m",
        "adjusted ebit $151m",
        "adjusted fcf $49m",
        "buybacks $87m",
        "2026 guidance raised",
        "business award",
        "2025-q4 - quarter notes",
        "q4 net sales $891m",
        "q4 adjusted ebit $122m",
        "q4 adjusted fcf $139m",
        "2025 year buybacks $208m",
        "initial 2026 outlook",
    ):
        assert required in text


def test_gtx_quarter_notes_matches_peer_depth_and_section_pattern(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Quarter_Notes_UI"]
    text = _sheet_visible_text(gtx_wb, "Quarter_Notes_UI").lower()
    quarter_blocks = [title.lower() for title in _quarter_note_titles(ws)]

    assert len(quarter_blocks) >= 10
    assert int(ws.max_row or 0) >= 260
    assert len(_non_empty_row_indices(ws, max_col=15)) >= 220
    assert "guidance / promise interpretation" in text
    assert "model mapping / double-count guardrails" in text
    for expected in (
        "2026-q1 - quarter notes",
        "2025-q4 - quarter notes",
        "2025-q3 - quarter notes",
        "2025-q2 - quarter notes",
        "2025-q1 - quarter notes",
        "2024-q4 - quarter notes",
        "2024-q3 - quarter notes",
        "2024-q2 - quarter notes",
        "2024-q1 - quarter notes",
        "2023-q4 - quarter notes",
    ):
        assert expected in quarter_blocks


def test_gtx_quarter_notes_uses_peer_style_title_and_quarter_blocks(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Quarter_Notes_UI"]
        merged = {str(rng) for rng in ws.merged_cells.ranges}
        quarter_rows = [
            rr
            for rr in range(1, int(ws.max_row or 0) + 1)
            if str(ws.cell(rr, 1).value or "").strip().lower().endswith("- quarter notes")
        ]

        assert str(ws["A1"].value or "").strip() == "2026-Q1 - Quarter Notes"
        assert str(ws["A2"].value or "").strip() == "Quarter read"
        assert ws.freeze_panes == "A2"
        assert any(str(rng).startswith("A1:") for rng in merged)
        assert len(quarter_rows) >= 10
        for expected in (
            "2026-Q1",
            "2025-Q4",
            "2025-Q3",
            "2025-Q2",
            "2025-Q1",
            "2024-Q4",
            "2024-Q3",
            "2024-Q2",
            "2024-Q1",
            "2023-Q4",
        ):
            assert any(str(ws.cell(rr, 1).value or "").startswith(expected) for rr in quarter_rows)
        assert quarter_rows[0] == 1
        for rr in quarter_rows:
            assert str(ws.cell(rr, 1).fill.fgColor.rgb or "").upper() in {
                "FF5B9BD5",
                "005B9BD5",
                "FF6FA8DC",
                "006FA8DC",
                "FF1F4E78",
                "001F4E78",
            }
            if rr + 5 <= int(ws.max_row or 0):
                fills = [
                    str(ws.cell(rr + offset, 1).fill.fgColor.rgb or "").upper()
                    for offset in (3, 4, 5)
                ]
                assert len(set(fills)) >= 2


def test_gtx_quarter_notes_top_block_uses_peer_row_height_template(
    gtx_wb: openpyxl.Workbook,
) -> None:
    """Missing source depth should not create a smaller GTX-specific top block."""
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Quarter_Notes_UI"]

        assert str(ws["A1"].value or "").strip() == "2026-Q1 - Quarter Notes"
        assert str(ws["A2"].value or "").strip() == "Quarter read"
        assert ws.freeze_panes == "A2"
        assert [
            float(ws.row_dimensions[row_idx].height or 0.0)
            for row_idx in range(1, 8)
        ] == pytest.approx([24.0, 25.0, 44.0, 44.0, 44.0, 30.0, 10.0])


def test_gtx_quarter_notes_uses_exact_peer_fills_and_body_heights(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Quarter_Notes_UI"]

        assert _rgb(ws["A1"]) == "005B9BD5"
        assert _rgb(ws["A2"]) == "00DDEBF7"
        assert _rgb(ws["A3"]) == "00F7FBFF"
        assert _rgb(ws["A4"]) == "00EDF4FB"
        assert _rgb(ws["A8"]) == "00DDEBF7"
        assert _rgb(ws["A9"]) == "00EAF3F8"

        for rr in (10, 11, 12, 13):
            assert float(ws.row_dimensions[rr].height or 0) == pytest.approx(48.0)
        assert float(ws.row_dimensions[8].height or 0) == pytest.approx(25.0)
        assert float(ws["A3"].font.sz or 0) == pytest.approx(14.0)


def test_gtx_quarter_notes_has_required_quarter_read_labels(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Quarter_Notes_UI"]
    required_quarters = {
        "2026-Q1 - Quarter Notes",
        "2025-Q4 - Quarter Notes",
        "2025-Q3 - Quarter Notes",
        "2025-Q2 - Quarter Notes",
        "2025-Q1 - Quarter Notes",
        "2024-Q4 - Quarter Notes",
        "2024-Q3 - Quarter Notes",
        "2024-Q2 - Quarter Notes",
        "2024-Q1 - Quarter Notes",
        "2023-Q4 - Quarter Notes",
    }
    quarter_rows = {
        str(ws.cell(rr, 1).value or "").strip(): rr
        for rr in range(1, int(ws.max_row or 0) + 1)
        if str(ws.cell(rr, 1).value or "").strip() in required_quarters
    }

    assert set(quarter_rows) == required_quarters
    for title, rr in quarter_rows.items():
        nearby_labels = {
            str(ws.cell(row_idx, 1).value or "").strip()
            for row_idx in range(rr, min(rr + 8, int(ws.max_row or 0)) + 1)
        }
        assert {"Quarter read", "Model read", "What changed", "Watch next", "Key caveat"}.issubset(
            nearby_labels
        ), title


def test_gtx_quarter_notes_key_development_layout_uses_peer_merges(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Quarter_Notes_UI"]
        merged = {str(rng) for rng in ws.merged_cells.ranges}
        header_rows = [
            rr
            for rr in range(1, int(ws.max_row or 0) + 1)
            if str(ws.cell(rr, 1).value or "").strip() == "Theme"
        ]

        assert header_rows
        for rr in header_rows[:5]:
            assert f"A{rr}:B{rr}" in merged
            assert f"C{rr}:E{rr}" in merged
            assert f"F{rr}:G{rr}" in merged
            assert f"H{rr}:L{rr}" in merged
            assert f"M{rr}:O{rr}" in merged
            assert str(ws.cell(rr, 13).value or "").strip() == "Source / confidence"
            central_headers = [
                str(ws.cell(rr, cc).value or "").strip().lower()
                for cc in range(1, 13)
            ]
            assert not any("source" in value or "confidence" in value for value in central_headers)

            first_body_row = rr + 1
            assert f"A{first_body_row}:B{first_body_row}" in merged
            assert f"C{first_body_row}:E{first_body_row}" in merged
            assert f"F{first_body_row}:G{first_body_row}" in merged
            assert f"H{first_body_row}:L{first_body_row}" in merged
            assert f"M{first_body_row}:O{first_body_row}" in merged


def test_gtx_quarter_notes_uses_peer_spacing_row_heights_and_source_width(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Quarter_Notes_UI"]
        quarter_rows = [
            rr
            for rr in range(1, int(ws.max_row or 0) + 1)
            if str(ws.cell(rr, 1).value or "").strip().endswith(" - Quarter Notes")
        ]
        blank_spacers = [
            rr
            for rr in range(2, int(ws.max_row or 0))
            if not any(str(ws.cell(rr, cc).value or "").strip() for cc in range(1, 16))
            and any(str(ws.cell(rr - 1, cc).value or "").strip() for cc in range(1, 16))
            and any(str(ws.cell(rr + 1, cc).value or "").strip() for cc in range(1, 16))
        ]
        theme_rows = [
            rr
            for rr in range(1, int(ws.max_row or 0) + 1)
            if str(ws.cell(rr, 1).value or "").strip() == "Theme"
        ]

        assert [str(ws.cell(rr, 1).value or "").strip() for rr in quarter_rows[:5]] == [
            "2026-Q1 - Quarter Notes",
            "2025-Q4 - Quarter Notes",
            "2025-Q3 - Quarter Notes",
            "2025-Q2 - Quarter Notes",
            "2025-Q1 - Quarter Notes",
        ]
        assert len(quarter_rows) >= 10
        assert blank_spacers
        assert all(float(ws.row_dimensions[rr].height or 0.0) <= 12.0 for rr in blank_spacers[:10])
        assert all(float(ws.row_dimensions[rr].height or 0.0) <= 30.0 for rr in quarter_rows[:5])
        assert all(float(ws.row_dimensions[rr + 1].height or 0.0) <= 30.0 for rr in quarter_rows[:5])
        assert max(float(ws.row_dimensions[rr].height or 0.0) for rr in range(1, int(ws.max_row or 0) + 1)) <= 62.5
        assert float(ws.column_dimensions["O"].width or 0.0) >= 40.0

        font_counts = _font_size_counts(ws, max_rows=400, max_cols=15)
        assert font_counts[13.0] + font_counts[14.0] >= 200
        assert font_counts[11.0] <= 10


def test_gtx_quarter_notes_core_reading_columns_are_not_clipped_by_long_unmerged_text(
    gtx_wb: openpyxl.Workbook,
) -> None:
    with _gtx_style_workbook() as full_wb:
        ws = full_wb["Quarter_Notes_UI"]
        merged_cells: set[tuple[int, int]] = set()
        for rng in ws.merged_cells.ranges:
            for rr in range(rng.min_row, rng.max_row + 1):
                for cc in range(rng.min_col, rng.max_col + 1):
                    merged_cells.add((rr, cc))

        offenders: list[tuple[str, int, int, int]] = []
        for row in ws.iter_rows(min_col=1, max_col=12):
            for cell in row:
                value = str(cell.value or "").strip()
                if not value or (cell.row, cell.column) in merged_cells:
                    continue
                if len(value) > 110:
                    offenders.append((cell.coordinate, cell.row, cell.column, len(value)))

        assert offenders == []


def test_gtx_history_q4_2025_operating_income_and_diluted_shares(
    gtx_wb: openpyxl.Workbook,
) -> None:
    q4_2025 = dt.date(2025, 12, 31)

    assert _history_value(gtx_wb, q4_2025, "op_income") == pytest.approx(
        119_000_000.0,
        abs=1_000_000.0,
    )
    assert _history_value(gtx_wb, q4_2025, "shares_diluted") == pytest.approx(
        197_514_000.0,
        abs=1_000_000.0,
    )


def test_gtx_older_q4_operating_income_is_not_impossible_vs_gross_profit(
    gtx_wb: openpyxl.Workbook,
) -> None:
    for quarter in (dt.date(2022, 12, 31), dt.date(2023, 12, 31)):
        gross_profit = _history_value(gtx_wb, quarter, "gross_profit")
        operating_income = _history_value(gtx_wb, quarter, "op_income")

        assert gross_profit is not None, f"GTX {quarter} gross profit should remain source-backed"
        assert operating_income is None or operating_income <= gross_profit, (
            f"GTX {quarter} operating income should be quarterized or blank, "
            f"not an impossible FY/YTD artifact above gross profit"
        )


def test_gtx_nongaap_credibility_does_not_show_impossible_q4_gaap_ebit(
    gtx_wb: openpyxl.Workbook,
) -> None:
    assert "NonGAAP_Credibility" in gtx_wb.sheetnames
    ws = gtx_wb["NonGAAP_Credibility"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    quarter_idx = headers.index("quarter")
    gaap_idx = headers.index("gaap_ebit")
    revenue_idx = headers.index("revenue")

    bad: list[tuple[Any, Any, Any]] = []
    impossible_values = {345_000_000, 388_000_000, 347_000_000, 353_000_000}
    for row in rows[1:]:
        quarter = row[quarter_idx]
        if not str(quarter).startswith(("2022-12-31", "2023-12-31", "2024-12-31", "2025-12-31")):
            continue
        gaap_ebit = row[gaap_idx]
        revenue = row[revenue_idx]
        if gaap_ebit in impossible_values:
            bad.append((quarter, gaap_ebit, revenue))

    assert bad == []


def test_gtx_nongaap_credibility_excludes_pre_package_local_fallback_rows(
    gtx_wb: openpyxl.Workbook,
) -> None:
    assert "NonGAAP_Credibility" in gtx_wb.sheetnames
    ws = gtx_wb["NonGAAP_Credibility"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    quarter_idx = headers.index("quarter")
    fallback_idx = headers.index("qa_fallback_source")
    snippet_idx = headers.index("evidence_snippet")

    fallback_rows: list[tuple[int, Any, Any]] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        quarter = row[quarter_idx]
        quarter_txt = str(quarter or "")
        fallback_txt = str(row[fallback_idx] or "").strip().lower()
        snippet_txt = str(row[snippet_idx] or "").strip().lower()
        if not quarter_txt.startswith(("2020-", "2021-", "2022-03-31")):
            continue
        if fallback_txt or "local fallback" in snippet_txt:
            fallback_rows.append((row_idx, quarter, row[snippet_idx]))

    assert fallback_rows == []


def test_gtx_adjusted_metrics_filters_low_confidence_local_fallback_rows(
    gtx_wb: openpyxl.Workbook,
) -> None:
    assert "Adjusted_Metrics" in gtx_wb.sheetnames
    ws = gtx_wb["Adjusted_Metrics"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    confidence_idx = headers.index("confidence")
    source_type_idx = headers.index("source_type")

    low_rows = [
        row
        for row in rows[1:]
        if str(row[confidence_idx] or "").strip().lower() == "low"
        or str(row[source_type_idx] or "").strip().lower() == "earnings_deck"
    ]

    assert low_rows == []


def test_gtx_q1_2026_fcf_qa_excludes_may_18_special_event_source(
    gtx_wb: openpyxl.Workbook,
) -> None:
    assert _history_value(gtx_wb, dt.date(2026, 3, 31), "cfo") == pytest.approx(
        98_000_000.0,
        abs=1_000_000.0,
    )
    assert _history_value(gtx_wb, dt.date(2026, 3, 31), "capex") == pytest.approx(
        29_000_000.0,
        abs=1_000_000.0,
    )

    qa_text = _sheet_text(gtx_wb, "QA_Log")
    fcf_qa_lines = [
        line
        for line in qa_text.splitlines()
        if "FCF (Q)" in line or "free cash flow" in line.lower()
    ]
    contaminated = [
        line
        for line in fcf_qa_lines
        if any(marker.lower() in line.lower() for marker in GTX_SPECIAL_EVENT_MARKERS)
    ]

    assert contaminated == []
    assert not any("extracted quarter text=$4.0m" in line for line in fcf_qa_lines)


def test_gtx_q1_2026_fcf_qa_labels_gaap_vs_adjusted_fcf_definition_difference(
    gtx_wb: openpyxl.Workbook,
) -> None:
    qa_text = _sheet_text(gtx_wb, "QA_Log")
    fcf_qa_lines = [
        line
        for line in qa_text.splitlines()
        if "FCF (Q)" in line or "free cash flow" in line.lower()
    ]

    assert any("CFO-capex" in line and "company-defined" in line for line in fcf_qa_lines)
    assert any("definition mismatch" in line.lower() for line in fcf_qa_lines)
    assert not any("likely conflicting extraction or source mismatch" in line for line in fcf_qa_lines)


def test_gtx_may_18_debt_event_does_not_contaminate_reported_history_or_qa(
    gtx_wb: openpyxl.Workbook,
) -> None:
    reported_surfaces = ("History_Q", "Adjusted_Metrics", "QA_Log")
    contaminated: list[tuple[str, str]] = []
    for sheet_name in reported_surfaces:
        if sheet_name not in gtx_wb.sheetnames:
            continue
        text = _sheet_text(gtx_wb, sheet_name)
        for marker in GTX_SPECIAL_EVENT_MARKERS:
            if marker.lower() in text.lower():
                contaminated.append((sheet_name, marker))

    assert contaminated == []


def test_gtx_may_2026_debt_repayment_is_labeled_post_quarter_event_context(
    gtx_wb: openpyxl.Workbook,
) -> None:
    text = _sheet_visible_text(gtx_wb, "Promise_Progress_UI")
    debt_event_lines = [
        line
        for line in text.splitlines()
        if "$50.0m disclosed" in line or "$50m" in line
    ]

    assert debt_event_lines, "Expected visible GTX debt repayment/event row"
    assert all(
        any(marker in line.lower() for marker in ("post-quarter", "event-context", "pro-forma", "pro forma"))
        for line in debt_event_lines
    )


def test_gtx_may_2026_debt_event_promise_row_is_concise(
    gtx_wb: openpyxl.Workbook,
) -> None:
    ws = gtx_wb["Promise_Progress_UI"]
    debt_rows: list[tuple[Any, ...]] = []
    for row in ws.iter_rows(values_only=True):
        row_text = " | ".join(str(value or "") for value in row)
        if "Debt reduction" in row_text and ("$50.0m" in row_text or "$50m" in row_text):
            debt_rows.append(row)

    assert debt_rows, "Expected GTX post-quarter debt-event row"
    for row in debt_rows:
        row_text = " | ".join(str(value or "") for value in row)
        assert re.search(r"\b(post[- ]quarter|event[- ]context|pro[- ]forma|pro forma)\b", row_text, re.I)
        assert len(row_text) < 420
        assert "Acquisition and divestiture expenses" not in row_text
        assert "Full year 2025 outlook" not in row_text


def test_gtx_debt_tranches_latest_omits_spurious_tiny_other_row(gtx_wb: openpyxl.Workbook) -> None:
    assert "Debt_Tranches_Latest" in gtx_wb.sheetnames
    ws = gtx_wb["Debt_Tranches_Latest"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    name_idx = headers.index("tranche_name")
    amount_idx = headers.index("amount_principal")

    tiny_other_rows = [
        row
        for row in rows[1:]
        if str(row[name_idx] or "").strip().lower() == "other"
        and float(row[amount_idx] or 0.0) < 5_000_000.0
    ]

    assert tiny_other_rows == []


def test_gtx_debt_tranches_latest_does_not_misread_term_facility_name_as_2025_maturity(
    gtx_wb: openpyxl.Workbook,
) -> None:
    assert "Debt_Tranches_Latest" in gtx_wb.sheetnames
    ws = gtx_wb["Debt_Tranches_Latest"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    name_idx = headers.index("tranche_name")
    maturity_year_idx = headers.index("maturity_year")
    maturity_display_idx = headers.index("maturity_display")

    term_rows = [
        row
        for row in rows[1:]
        if str(row[name_idx] or "").strip() == "2025 Dollar Term Facility"
    ]

    assert len(term_rows) == 1
    term_row = term_rows[0]
    assert term_row[maturity_year_idx] != 2025
    assert str(term_row[maturity_display_idx] or "").strip() != "2025"
