from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_DIR = Path(
    os.environ.get(
        "STOCK_MODEL_WORKBOOK_DIR",
        str(REPO_ROOT / "StockModelData" / "outputs" / "Excel stock models"),
    )
)
TICKERS = ("PBI", "GPRE", "ANF")


def _load_workbook(ticker: str, *, data_only: bool = False) -> Any:
    xlsx_path = WORKBOOK_DIR / f"{ticker}_model.xlsx"
    xlsm_path = WORKBOOK_DIR / f"{ticker}_model.xlsm"
    path = xlsx_path
    if xlsm_path.exists() and (not xlsx_path.exists() or xlsm_path.stat().st_mtime >= xlsx_path.stat().st_mtime):
        path = xlsm_path
    if not path.exists():
        pytest.skip(f"{path} is not available for hidden value UI tests")
    return load_workbook(path, data_only=data_only, read_only=False)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _header_map(ws: Any) -> dict[str, int]:
    return {_text(ws.cell(1, cc).value): cc for cc in range(1, int(ws.max_column or 0) + 1) if _text(ws.cell(1, cc).value)}


def _row_by_label(ws: Any, label: str) -> int:
    for rr in range(1, int(ws.max_row or 0) + 1):
        if _text(ws.cell(rr, 1).value) == label:
            return rr
    raise AssertionError(f"{ws.title}: could not find {label!r}")


def test_hidden_value_flags_have_explicit_trigger_and_nonblank_scores() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker, data_only=True)
        wb_formula = _load_workbook(ticker, data_only=False)
        try:
            assert "Hidden_Value_Flags" in wb.sheetnames, f"{ticker}: missing Hidden_Value_Flags"
            ws = wb["Hidden_Value_Flags"]
            ws_formula = wb_formula["Hidden_Value_Flags"]
            headers = _header_map(ws)
            for required in {"flag_code", "title", "score", "severity", "visible_support", "triggered"}:
                assert required in headers, f"{ticker}: Hidden_Value_Flags missing {required!r}"
            code_col = headers["flag_code"]
            score_col = headers["score"]
            triggered_col = headers["triggered"]
            seen_nontriggered = False
            for rr in range(2, int(ws.max_row or 0) + 1):
                flag_code = _text(ws.cell(rr, code_col).value)
                title = _text(ws.cell(rr, headers["title"]).value)
                if not flag_code and not title:
                    continue
                score = ws.cell(rr, score_col).value
                triggered = ws.cell(rr, triggered_col).value
                score_formula = _text(ws_formula.cell(rr, score_col).value)
                triggered_formula = _text(ws_formula.cell(rr, triggered_col).value)
                if flag_code in {"C", "E"} and score in (None, ""):
                    assert "Hidden_Value_Audit" in score_formula, (
                        f"{ticker}: row {rr} price-linked score should be formula-backed, got {score_formula!r}"
                    )
                else:
                    assert score not in (None, ""), f"{ticker}: row {rr} has visible flag text but blank score"
                    assert isinstance(score, (int, float)), f"{ticker}: row {rr} score should be numeric, got {score!r}"
                if flag_code in {"C", "E"} and triggered not in (0, 1, False, True):
                    assert "Hidden_Value_Audit" in triggered_formula, (
                        f"{ticker}: row {rr} price-linked trigger should be formula-backed, got {triggered_formula!r}"
                    )
                    continue
                assert triggered in (0, 1, False, True), f"{ticker}: row {rr} triggered should be explicit 0/1"
                if not bool(triggered):
                    seen_nontriggered = True
                else:
                    assert not seen_nontriggered, f"{ticker}: triggered rows should sort before non-triggered audit rows"
                    assert float(score) >= 1.0, f"{ticker}: triggered row {rr} should have a positive score"
        finally:
            wb.close()
            wb_formula.close()


def test_valuation_hidden_flags_gate_on_triggered_not_blank_score() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker, data_only=False)
        try:
            ws = wb["Valuation"]
            flags_header_row = _row_by_label(ws, "Hidden value flags")
            helper_formula = _text(ws.cell(flags_header_row + 2, 35).value)
            normalized = helper_formula.replace("'", "")
            assert "Hidden_Value_Flags!$L$2:$L$100" in normalized, (
                f"{ticker}: hidden flag helper should scan explicit triggered column, got {helper_formula!r}"
            )
            assert "_xludf" not in normalized.lower() and "aggregate(" not in normalized.lower(), (
                f"{ticker}: hidden flag helper should use compatible MATCH/INDEX formulas, got {helper_formula!r}"
            )
            assert "Hidden_Value_Flags!$D$2:$D$100" not in normalized and "Hidden_Value_Flags!$D$2" not in normalized, (
                f"{ticker}: helper should not treat blank score as trigger state, got {helper_formula!r}"
            )
        finally:
            wb.close()


def test_valuation_hidden_flags_have_cached_display_values() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker, data_only=True)
        try:
            ws_val = wb["Valuation"]
            ws_hvf = wb["Hidden_Value_Flags"]
            headers = _header_map(ws_hvf)
            triggered_col = headers["triggered"]
            title_col = headers["title"]
            score_col = headers["score"]
            severity_col = headers["severity"]
            support_col = headers["visible_support"]
            triggered_rows = [
                rr
                for rr in range(2, int(ws_hvf.max_row or 0) + 1)
                if ws_hvf.cell(rr, triggered_col).value in (1, True)
            ]

            flags_header_row = _row_by_label(ws_val, "Hidden value flags")
            first_display_row = flags_header_row + 2
            if triggered_rows:
                first_triggered = triggered_rows[0]
                assert _text(ws_val.cell(first_display_row, 1).value) == "Flag 1", (
                    f"{ticker}: Valuation hidden flag display should not depend on an empty formula cache"
                )
                assert _text(ws_val.cell(first_display_row, 2).value) == _text(
                    ws_hvf.cell(first_triggered, title_col).value
                )
                assert ws_val.cell(first_display_row, 6).value == ws_hvf.cell(first_triggered, score_col).value
                assert _text(ws_val.cell(first_display_row, 7).value) == _text(
                    ws_hvf.cell(first_triggered, severity_col).value
                )
                assert _text(ws_val.cell(first_display_row, 8).value) == _text(
                    ws_hvf.cell(first_triggered, support_col).value
                )
            else:
                assert _text(ws_val.cell(first_display_row, 1).value) == "No triggered flags"
                assert _text(ws_val.cell(first_display_row, 2).value) == "No scored hidden-value flags currently triggered"
                assert _text(ws_val.cell(first_display_row, 7).value) == "Info"
        finally:
            wb.close()


def test_price_linked_hidden_value_flags_flow_from_audit_to_visible_candidates() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker, data_only=False)
        try:
            assert "Hidden_Value_Flags" in wb.sheetnames
            assert "Hidden_Value_Audit" in wb.sheetnames
            ws = wb["Hidden_Value_Flags"]
            headers = _header_map(ws)
            code_col = headers["flag_code"]
            score_col = headers["score"]
            trigger_col = headers["triggered"]
            price_linked = {}
            for rr in range(2, int(ws.max_row or 0) + 1):
                code = _text(ws.cell(rr, code_col).value)
                if code in {"C", "E"}:
                    price_linked[code] = {
                        "score": _text(ws.cell(rr, score_col).value),
                        "triggered": _text(ws.cell(rr, trigger_col).value),
                    }

            assert set(price_linked) == {"C", "E"}, f"{ticker}: expected C/E price-linked audit candidates"
            for code, values in price_linked.items():
                assert "Hidden_Value_Audit" in values["score"], (
                    f"{ticker}: {code} score should be formula-linked to audit output, got {values['score']!r}"
                )
                assert "Hidden_Value_Audit" in values["triggered"], (
                    f"{ticker}: {code} trigger should be formula-linked to audit output, got {values['triggered']!r}"
                )
        finally:
            wb.close()


def test_valuation_hidden_value_panel_uses_compact_styled_rows() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker, data_only=False)
        wb_values = _load_workbook(ticker, data_only=True)
        try:
            ws = wb["Valuation"]
            flags_ws = wb_values["Hidden_Value_Flags"]
            headers = _header_map(flags_ws)
            triggered_col = headers["triggered"]
            code_col = headers["flag_code"]
            visible_codes = set()
            for rr in range(2, int(flags_ws.max_row or 0) + 1):
                code = _text(flags_ws.cell(rr, code_col).value)
                if flags_ws.cell(rr, triggered_col).value in (1, True):
                    visible_codes.add(code or f"row:{rr}")
                if code in {"C", "E"}:
                    visible_codes.add(code)

            flags_header_row = _row_by_label(ws, "Hidden value flags")
            visible_count = max(1, min(5, len(visible_codes)))
            first_unused_flag_row = flags_header_row + 2 + visible_count
            if first_unused_flag_row <= flags_header_row + 8:
                assert not _text(ws.cell(first_unused_flag_row, 35).value), (
                    f"{ticker}: hidden flag UI should not leave old helper/formula rows below visible flags"
                )
                assert not _text(ws.cell(first_unused_flag_row, 1).value), (
                    f"{ticker}: hidden flag UI should not leave blank bordered flag rows"
                )
                assert ws.cell(first_unused_flag_row, 1).border.left.style is None, (
                    f"{ticker}: first unused hidden flag row should be visually blank"
                )

            panel_header_row = flags_header_row
            assert any(
                m.min_row == panel_header_row
                and m.max_row == panel_header_row
                and m.min_col == 14
                and m.max_col == 18
                for m in ws.merged_cells.ranges
            ), f"{ticker}: Hidden Value Panel header should be merged across N:R"
            for rr in range(panel_header_row + 1, panel_header_row + 7):
                assert any(
                    m.min_row == rr and m.max_row == rr and m.min_col == 14 and m.max_col == 17
                    for m in ws.merged_cells.ranges
                ), f"{ticker}: Hidden Value Panel label row {rr} should be merged across N:Q"
                for cc in (14, 18):
                    cell = ws.cell(rr, cc)
                    assert cell.border.left.style == "thin", f"{ticker}: panel row {rr} col {cc} missing border"
                    assert cell.fill.fill_type == "solid", f"{ticker}: panel row {rr} col {cc} missing body fill"
        finally:
            wb.close()
            wb_values.close()
