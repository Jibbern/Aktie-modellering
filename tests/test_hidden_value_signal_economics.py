from __future__ import annotations

import copy
import json
import random
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import pytest
from openpyxl import load_workbook

from pbi_xbrl.hidden_value_signal_economics import (
    HIDDEN_VALUE_PLAN_SCHEMA,
    ResolvedMetric,
    _evaluate_signal,
    evaluate_hidden_value_signals,
    load_hidden_value_signal_contract,
)
from pbi_xbrl.json_schema_validation import load_json_strict, validate_json_schema
from pbi_xbrl.workbook_modules import load_workbook_module_manifest


PERIODS = [f"{year}-Q{quarter}" for year in (2024, 2025) for quarter in range(1, 5)]


def _source_row(period: str, metric: str, value: float, unit: str) -> dict[str, Any]:
    return {
        "period": period,
        "metric": metric,
        "value": value,
        "unit": unit,
        "status": "source_backed",
        "source_ref": f"fixture:{metric}:{period}",
    }


def _economic_package() -> dict[str, Any]:
    series = {
        "revenue": ([100.0] * 4 + [150.0] * 4, "$m"),
        "operating_income": ([10.0] * 4 + [13.0] * 4, "$m"),
        "base_ebitda": ([15.0] * 4 + [18.75] * 4, "$m"),
        "adjusted_ebitda": ([20.0] * 4 + [37.5] * 4, "$m"),
        "operating_cash_flow": ([20.0] * 4 + [40.0] * 4, "$m"),
        "capital_expenditures": ([5.0] * 8, "$m"),
        "interest_expense": ([4.0] * 4 + [3.0] * 4, "$m"),
        "shares_outstanding": ([100.0] * 4 + [99.5, 99.0, 98.0, 97.0], "m shares"),
        "market_cap": ([700.0] * 4 + [500.0] * 4, "$m"),
        "net_debt": ([300.0] * 4 + [280.0, 250.0, 225.0, 200.0], "$m"),
        "dividend_per_share": ([0.10] * 4 + [0.11, 0.11, 0.12, 0.12], "$/share"),
    }
    rows = [
        _source_row(period, metric, values[index], unit)
        for metric, (values, unit) in series.items()
        for index, period in enumerate(PERIODS)
    ]
    return {
        "ticker_metadata": {"ticker": "TEST"},
        "calculation_history": {"quarterly_items": rows},
        "annual_financials": {"rows": []},
        "valuation_inputs": {
            "price": {
                "value": None,
                "unit": "$/share",
                "period": "2025-Q4",
                "status": "missing_source",
                "source_ref": "fixture:price:missing",
            }
        },
    }


def _candidate(plan: Any, signal_id: str) -> Any:
    return next(row for row in plan.candidates if row.signal_id == signal_id)


def _candidate_from_exact_metrics(signal_id: str, values: dict[str, float]) -> Any:
    contract = load_hidden_value_signal_contract()
    signals = {row["signal_id"]: row for row in contract["signals"]}
    metrics = {row["metric_id"]: row for row in contract["metric_resolvers"]}
    signal = signals[signal_id]
    resolved = {
        metric_id: ResolvedMetric(
            metric_id,
            value,
            metrics[metric_id]["unit"],
            "2025-Q4",
            metrics[metric_id]["period_role"],
            "derived_calculated",
            "independent_boundary_fixture",
            (f"fixture:{signal_id}:{metric_id}",),
            (),
        )
        for metric_id, value in values.items()
    }
    resolver = SimpleNamespace(
        latest_period="2025-Q4",
        resolve=lambda metric_id, period: resolved[metric_id],
    )
    return _evaluate_signal(
        signal,
        resolver,
        "anf",
        set(signal["required_modules"]),
        set(),
        contract,
    )


def _rows(package: dict[str, Any], metric: str) -> list[dict[str, Any]]:
    return [
        row
        for row in package["calculation_history"]["quarterly_items"]
        if row["metric"] == metric
    ]


def _set_value(package: dict[str, Any], metric: str, period: str, value: float) -> None:
    next(row for row in _rows(package, metric) if row["period"] == period)["value"] = value


def _set_series(package: dict[str, Any], metric: str, values: list[float]) -> None:
    by_period = {row["period"]: row for row in _rows(package, metric)}
    for period, value in zip(PERIODS, values):
        by_period[period]["value"] = value


def _set_price(package: dict[str, Any], value: float) -> None:
    package["valuation_inputs"]["price"] = {
        "value": value,
        "unit": "$/share",
        "period": "2025-Q4",
        "status": "source_backed",
        "source_ref": "fixture:price:2025-Q4",
    }


def _workbook_path(ticker: str) -> Path:
    aktier_root = next(parent for parent in Path(__file__).resolve().parents if parent.name == "Aktier")
    root = aktier_root / "StockModelData" / "outputs" / "Excel stock models"
    xlsx = root / f"{ticker}_model.xlsx"
    xlsm = root / f"{ticker}_model.xlsm"
    if xlsm.exists() and (not xlsx.exists() or xlsm.stat().st_mtime >= xlsx.stat().st_mtime):
        return xlsm
    return xlsx


def _legacy_flags(ticker: str) -> dict[str, tuple[Any, Any]]:
    path = _workbook_path(ticker)
    if not path.exists():
        pytest.skip(f"Legacy oracle {path} is unavailable")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Hidden_Value_Flags"]
        headers = {str(sheet.cell(1, column).value or "").strip(): column for column in range(1, sheet.max_column + 1)}
        result: dict[str, tuple[Any, Any]] = {}
        for row in range(2, sheet.max_row + 1):
            code = str(sheet.cell(row, headers["flag_code"]).value or "").strip()
            if code:
                result[code] = (
                    sheet.cell(row, headers["score"]).value,
                    sheet.cell(row, headers["triggered"]).value,
                )
        return result
    finally:
        workbook.close()


def test_exact_synthetic_fixture_triggers_all_a_to_g() -> None:
    plan = evaluate_hidden_value_signals(_economic_package(), profile_id="anf", ticker="TEST")

    assert [row.signal_id for row in plan.candidates] == list("ABCDEFG")
    assert {row.signal_id for row in plan.candidates if row.triggered} == set("ABCDEFG")
    assert all(row.state == "triggered" for row in plan.candidates)
    assert len(plan.flags_projection) == 7
    assert all(row["triggered"] is True for row in plan.flags_projection)
    assert all(row.score is not None and 0 <= row.score <= 100 for row in plan.candidates)
    assert all(row.source_refs and row.evidence_ids for row in plan.candidates)
    assert {row.signal_id: row.score for row in plan.candidates} == {
        "A": 41,
        "B": 82,
        "C": 92,
        "D": 71,
        "E": 61,
        "F": 49,
        "G": 69,
    }
    assert [row["signal_id"] for row in plan.flags_projection] == ["C", "B", "D", "G", "E", "F", "A"]


@pytest.mark.parametrize(
    ("signal_id", "mutation"),
    [
        ("A", lambda package: _set_value(package, "shares_outstanding", "2025-Q4", 100.0)),
        ("B", lambda package: _set_series(package, "adjusted_ebitda", [20.0] * 8)),
        ("C", lambda package: _set_series(package, "market_cap", [2500.0] * 8)),
        ("D", lambda package: _set_value(package, "net_debt", "2025-Q4", 295.0)),
        ("E", lambda package: _set_series(package, "interest_expense", [4.0] * 4 + [15.0] * 4)),
        ("F", lambda package: _set_value(package, "shares_outstanding", "2025-Q4", 99.5)),
        ("G", lambda package: _set_series(package, "dividend_per_share", [0.10] * 8)),
    ],
)
def test_each_signal_has_an_independent_not_triggered_fixture(signal_id: str, mutation: Any) -> None:
    package = _economic_package()
    mutation(package)

    candidate = _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), signal_id)

    assert candidate.triggered is False
    assert candidate.state in {"not_triggered", "near_miss"}


def test_contract_locks_all_trigger_and_near_miss_threshold_boundaries() -> None:
    signals = {row["signal_id"]: row for row in load_hidden_value_signal_contract()["signals"]}
    trigger = {
        signal_id: [(row["metric_id"], row["operator"], row["threshold"]) for row in signal["trigger"]["predicates"]]
        for signal_id, signal in signals.items()
    }
    assert trigger == {
        "A": [("ebit_growth_yoy", "gt", 0.25), ("base_ebitda_growth_yoy", "gt", 0.20), ("shares_outstanding_yoy", "lte", -0.02)],
        "B": [("adjusted_ebitda_margin_ttm", "gte", 0.20), ("adjusted_ebitda_margin_yoy_bps", "gte", 200), ("adjusted_ebitda_margin_streak", "gte", 2)],
        "C": [("positive_fcf_ttm_observations", "gte", 1), ("fcf_to_ebit", "gte", 0.75), ("fcf_yield", "gte", 0.15)],
        "D": [("net_debt_reduction", "gte", 0.10), ("net_leverage", "lte", 3.0)],
        "E": [("interest_coverage", "gte", 3.0), ("fcf_yield", "gte", 0.20)],
        "F": [("shares_outstanding_yoy", "lte", -0.02)],
        "G": [("dividend_yield", "gte", 0.03), ("dividend_per_share_yoy", "gt", 0), ("dividend_per_share_qoq", "gt", 0)],
    }
    near = {
        signal_id: [(row["metric_id"], row["operator"], row["threshold"]) for row in signal["near_miss"]["predicates"]]
        for signal_id, signal in signals.items()
    }
    assert near["A"] == [("ebit_growth_yoy", "gte", 0.20), ("base_ebitda_growth_yoy", "gte", 0.15), ("shares_outstanding_yoy", "lte", -0.01)]
    assert near["B"] == [("adjusted_ebitda_margin_ttm", "gte", 0.18), ("adjusted_ebitda_margin_yoy_bps", "gte", 150), ("adjusted_ebitda_margin_streak", "gte", 1)]
    assert near["C"] == [("positive_fcf_ttm_observations", "gte", 1), ("fcf_to_ebit", "gte", 0.65), ("fcf_yield", "gte", 0.12)]
    assert near["D"] == [("net_debt_reduction", "gte", 0.05), ("net_leverage", "lte", 3.5)]
    assert near["E"] == [("interest_coverage", "gte", 2.5), ("fcf_yield", "gte", 0.15)]
    assert near["F"] == [("shares_outstanding_yoy", "lte", -0.01)]
    assert near["G"] == [("dividend_yield", "gte", 0.025), ("dividend_yield", "lt", 0.03)]


def test_only_signal_a_requires_two_trigger_predicates_for_near_miss() -> None:
    signals = {row["signal_id"]: row for row in load_hidden_value_signal_contract()["signals"]}

    assert signals["A"]["near_miss"]["minimum_trigger_predicates"] == 2
    assert "minimum_trigger_predicates" not in signals["B"]["near_miss"]
    assert "minimum_trigger_predicates" not in signals["C"]["near_miss"]


def test_b_and_c_exact_near_miss_boundaries_are_classified_as_near_miss() -> None:
    signal_b = _candidate_from_exact_metrics(
        "B",
        {
            "adjusted_ebitda_ttm": 100.0,
            "adjusted_ebitda_margin_ttm": 0.18,
            "adjusted_ebitda_margin_yoy_bps": 150.0,
            "adjusted_ebitda_margin_streak": 1.0,
            "revenue_ttm": 500.0,
        },
    )
    signal_c = _candidate_from_exact_metrics(
        "C",
        {
            "positive_fcf_ttm_observations": 1.0,
            "fcf_to_ebit": 0.65,
            "fcf_yield": 0.12,
            "ebit_ttm": 100.0,
            "market_cap": 1_000.0,
        },
    )

    assert signal_b.state == "near_miss"
    assert signal_b.triggered is False
    assert signal_c.state == "near_miss"
    assert signal_c.triggered is False


def test_b_and_c_immediately_outside_near_miss_boundaries_are_not_triggered() -> None:
    signal_b = _candidate_from_exact_metrics(
        "B",
        {
            "adjusted_ebitda_ttm": 100.0,
            "adjusted_ebitda_margin_ttm": 0.18 - 1e-9,
            "adjusted_ebitda_margin_yoy_bps": 150.0,
            "adjusted_ebitda_margin_streak": 1.0,
            "revenue_ttm": 500.0,
        },
    )
    signal_c = _candidate_from_exact_metrics(
        "C",
        {
            "positive_fcf_ttm_observations": 1.0,
            "fcf_to_ebit": 0.65 - 1e-9,
            "fcf_yield": 0.12,
            "ebit_ttm": 100.0,
            "market_cap": 1_000.0,
        },
    )

    assert signal_b.state == "not_triggered"
    assert signal_b.triggered is False
    assert signal_c.state == "not_triggered"
    assert signal_c.triggered is False


def test_f_and_g_exact_trigger_and_near_miss_boundaries() -> None:
    package = _economic_package()
    _set_value(package, "shares_outstanding", "2025-Q4", 98.0)
    assert _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "F").state == "triggered"

    package = _economic_package()
    _set_value(package, "shares_outstanding", "2025-Q4", 99.0)
    assert _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "F").state == "near_miss"

    package = _economic_package()
    _set_series(package, "dividend_per_share", [0.12] * 8)
    _set_price(package, 16.0)
    assert _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "G").state == "triggered"

    package = _economic_package()
    _set_series(package, "dividend_per_share", [0.10] * 8)
    _set_price(package, 16.0)
    assert _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "G").state == "near_miss"


def test_a_to_e_exact_trigger_boundaries_follow_economic_rules() -> None:
    package = _economic_package()
    _set_series(package, "operating_income", [10.0] * 4 + [12.5] * 4)
    _set_series(package, "base_ebitda", [15.0] * 4 + [18.0] * 4)
    _set_value(package, "shares_outstanding", "2025-Q4", 98.0)
    assert _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "A").state == "not_triggered"

    package = _economic_package()
    _set_series(package, "adjusted_ebitda", [18.0] * 4 + [30.0] * 4)
    assert _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "B").state == "triggered"

    package = _economic_package()
    _set_series(package, "operating_cash_flow", [20.0] * 4 + [14.75] * 4)
    _set_series(package, "capital_expenditures", [5.0] * 8)
    _set_series(package, "market_cap", [700.0] * 4 + [260.0] * 4)
    assert _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "C").state == "triggered"

    package = _economic_package()
    _set_series(package, "base_ebitda", [15.0] * 4 + [22.5] * 4)
    _set_value(package, "net_debt", "2025-Q4", 270.0)
    assert _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "D").state == "triggered"

    package = _economic_package()
    _set_series(package, "interest_expense", [4.0] * 4 + [52.0 / 12.0] * 4)
    _set_series(package, "market_cap", [700.0] * 8)
    assert _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "E").state == "triggered"


def test_signal_a_prior_base_must_strictly_exceed_materiality_floor() -> None:
    package = _economic_package()
    _set_series(package, "operating_income", [2.5] * 4 + [5.0] * 4)

    candidate = _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "A")

    assert candidate.triggered is False
    assert candidate.state == "not_triggered"
    assert "eligibility_failed:prior_ebit_material" in candidate.reasons


def test_signal_b_score_includes_the_streak_component() -> None:
    candidate = _candidate(evaluate_hidden_value_signals(_economic_package(), profile_id="anf"), "B")
    components = {row.component_id: row for row in candidate.component_scores}
    adjusted_margin = next(row for row in candidate.resolved_inputs if row.metric_id == "adjusted_ebitda_margin_ttm")

    assert candidate.score_denominator == 100
    assert components["margin_streak_score"].included_weight == 25
    assert components["margin_streak_score"].weighted_score is not None
    assert candidate.score == round(sum(float(row.weighted_score) for row in candidate.component_scores))
    assert set(adjusted_margin.formula_ids) == {
        "adjusted_ebitda_margin_ttm",
        "adjusted_ebitda_ttm",
        "revenue_ttm",
    }


def test_optional_fcf_per_share_component_is_reweighted_not_zero_filled() -> None:
    package = _economic_package()
    latest_cfo = next(row for row in _rows(package, "operating_cash_flow") if row["period"] == "2025-Q4")
    latest_cfo["status"] = "missing_source"
    latest_cfo["value"] = 0.0

    candidate = _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "F")

    assert candidate.triggered is True
    assert candidate.score_denominator == 70
    optional = next(row for row in candidate.component_scores if row.component_id == "fcf_per_share_score")
    assert optional.included_weight == 0
    assert optional.weighted_score is None
    shares = next(row for row in candidate.component_scores if row.component_id == "share_reduction_score")
    assert candidate.score == round(float(shares.normalized_score) * 100)


def test_share_reduction_and_fcf_per_share_fixture_scores_f_at_100() -> None:
    package = _economic_package()
    _set_value(package, "shares_outstanding", "2025-Q4", 85.0)

    candidate = _candidate(evaluate_hidden_value_signals(package, profile_id="pbi"), "F")

    assert candidate.triggered is True
    assert candidate.score == 100
    assert candidate.score_denominator == 100


def test_missing_status_unit_period_and_formula_inputs_fail_closed() -> None:
    missing_status = _economic_package()
    shares = next(row for row in _rows(missing_status, "shares_outstanding") if row["period"] == "2025-Q4")
    shares["status"] = "manual_review_required"
    shares["value"] = 0.0
    assert _candidate(evaluate_hidden_value_signals(missing_status, profile_id="anf"), "F").state == "insufficient_evidence"

    unit_mismatch = _economic_package()
    shares = next(row for row in _rows(unit_mismatch, "shares_outstanding") if row["period"] == "2025-Q4")
    shares["unit"] = "$m"
    assert _candidate(evaluate_hidden_value_signals(unit_mismatch, profile_id="anf"), "F").state == "invalid_input"

    missing_period = _economic_package()
    missing_period["calculation_history"]["quarterly_items"] = [
        row
        for row in missing_period["calculation_history"]["quarterly_items"]
        if not (row["metric"] == "operating_income" and row["period"] == "2025-Q3")
    ]
    assert _candidate(evaluate_hidden_value_signals(missing_period, profile_id="anf"), "A").state == "insufficient_evidence"

    missing_adjusted = _economic_package()
    adjusted = next(row for row in _rows(missing_adjusted, "adjusted_ebitda") if row["period"] == "2025-Q4")
    adjusted["status"] = "missing_source"
    adjusted["value"] = 0.0
    assert _candidate(evaluate_hidden_value_signals(missing_adjusted, profile_id="anf"), "B").state == "insufficient_evidence"


def test_explicit_missing_source_disposition_is_unavailable() -> None:
    package = _economic_package()
    package["calculation_history"]["quarterly_items"] = [
        row for row in package["calculation_history"]["quarterly_items"] if row["metric"] != "dividend_per_share"
    ]
    package["capital_returns"] = {
        "dividends": {
            "value": None,
            "status": "missing_source",
            "source_ref": "fixture:dividends:missing",
        }
    }

    candidate = _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "G")

    assert candidate.state == "unavailable"
    assert candidate.reasons == ("dividend_per_share:source_missing_source",)


def test_dimension_period_and_lineage_mismatches_fail_closed() -> None:
    wrong_dimension = _economic_package()
    shares = next(row for row in _rows(wrong_dimension, "shares_outstanding") if row["period"] == "2025-Q4")
    shares["dimension_id"] = "geography"
    shares["member"] = "Americas"
    candidate = _candidate(evaluate_hidden_value_signals(wrong_dimension, profile_id="anf"), "F")
    assert candidate.state == "insufficient_evidence"
    assert any("dimension_member_mismatch" in reason for reason in candidate.reasons)

    wrong_period = _economic_package()
    wrong_period["calculation_history"]["quarterly_items"] = [
        row for row in wrong_period["calculation_history"]["quarterly_items"] if row["metric"] != "market_cap"
    ]
    _set_price(wrong_period, 5.0)
    wrong_period["valuation_inputs"]["price"]["period"] = "2025-Q3"
    assert _candidate(evaluate_hidden_value_signals(wrong_period, profile_id="anf"), "C").state == "invalid_input"

    missing_lineage = _economic_package()
    missing_lineage["calculation_history"]["quarterly_items"] = [
        row for row in missing_lineage["calculation_history"]["quarterly_items"] if row["metric"] != "market_cap"
    ]
    _set_price(missing_lineage, 5.0)
    missing_lineage["valuation_inputs"]["price"]["source_ref"] = ""
    assert _candidate(evaluate_hidden_value_signals(missing_lineage, profile_id="anf"), "C").state == "insufficient_evidence"


def test_net_debt_requires_source_evidence_and_positive_prior_base() -> None:
    invalid_prior = _economic_package()
    _set_value(invalid_prior, "net_debt", "2024-Q4", -100.0)
    candidate = _candidate(evaluate_hidden_value_signals(invalid_prior, profile_id="anf"), "D")
    assert candidate.state == "invalid_input"
    assert candidate.score is None

    net_cash_current = _economic_package()
    _set_value(net_cash_current, "net_debt", "2025-Q4", -50.0)
    candidate = _candidate(evaluate_hidden_value_signals(net_cash_current, profile_id="anf"), "D")
    assert candidate.triggered is True
    assert next(row for row in candidate.resolved_inputs if row.metric_id == "net_debt").value == -50.0

    no_net_debt = _economic_package()
    no_net_debt["calculation_history"]["quarterly_items"] = [
        row for row in no_net_debt["calculation_history"]["quarterly_items"] if row["metric"] != "net_debt"
    ]
    assert _candidate(evaluate_hidden_value_signals(no_net_debt, profile_id="anf"), "D").state == "insufficient_evidence"


def test_market_cap_accepts_exact_price_times_point_in_time_shares_only() -> None:
    package = _economic_package()
    package["calculation_history"]["quarterly_items"] = [
        row for row in package["calculation_history"]["quarterly_items"] if row["metric"] != "market_cap"
    ]
    _set_price(package, 5.0)
    candidate = _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "C")
    market_cap = next(row for row in candidate.resolved_inputs if row.metric_id == "market_cap")
    assert market_cap.value == pytest.approx(485.0)
    assert market_cap.source_refs == ("fixture:price:2025-Q4", "fixture:shares_outstanding:2025-Q4")

    diluted_only = copy.deepcopy(package)
    diluted_only["calculation_history"]["quarterly_items"] = [
        row for row in diluted_only["calculation_history"]["quarterly_items"] if row["metric"] != "shares_outstanding"
    ]
    diluted_only["calculation_history"]["quarterly_items"].append(
        _source_row("2025-Q4", "diluted_shares", 97.0, "m shares")
    )
    assert _candidate(evaluate_hidden_value_signals(diluted_only, profile_id="anf"), "C").state == "insufficient_evidence"


def test_cash_interest_and_base_adjusted_fallbacks_are_not_used() -> None:
    package = _economic_package()
    package["calculation_history"]["quarterly_items"] = [
        row for row in package["calculation_history"]["quarterly_items"] if row["metric"] != "interest_expense"
    ]
    for period in PERIODS:
        package["calculation_history"]["quarterly_items"].append(_source_row(period, "interest_paid", 1.0, "$m"))
    assert _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "E").state == "insufficient_evidence"

    package = _economic_package()
    package["calculation_history"]["quarterly_items"] = [
        row for row in package["calculation_history"]["quarterly_items"] if row["metric"] != "adjusted_ebitda"
    ]
    assert _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "B").state == "insufficient_evidence"


def test_plan_and_all_projections_are_row_order_independent() -> None:
    package = _economic_package()
    shuffled = copy.deepcopy(package)
    random.Random(20260717).shuffle(shuffled["calculation_history"]["quarterly_items"])

    left = evaluate_hidden_value_signals(package, profile_id="anf").to_dict()
    right = evaluate_hidden_value_signals(shuffled, profile_id="anf").to_dict()

    assert left == right
    assert all(row["state"] == "triggered" for row in left["projections"]["flags"])
    assert len(left["projections"]["audit"]) == 7
    assert len(left["projections"]["flags"]) == 7


def test_generated_projection_schema_rejects_untyped_or_nontriggered_flag_rows() -> None:
    payload = evaluate_hidden_value_signals(_economic_package(), profile_id="anf").to_dict()
    schema = load_json_strict(HIDDEN_VALUE_PLAN_SCHEMA)
    assert validate_json_schema(payload, schema) == []

    extra_field = copy.deepcopy(payload)
    extra_field["projections"]["base"][0]["unexpected"] = True
    assert validate_json_schema(extra_field, schema)

    false_flag = copy.deepcopy(payload)
    false_flag["projections"]["flags"][0]["triggered"] = False
    assert validate_json_schema(false_flag, schema)


def test_recompute_projection_contains_every_predicate_and_score_component() -> None:
    contract = load_hidden_value_signal_contract()
    package = _economic_package()
    shares = next(row for row in _rows(package, "shares_outstanding") if row["period"] == "2025-Q4")
    shares["status"] = "missing_source"
    plan = evaluate_hidden_value_signals(package, profile_id="anf")

    expected = {
        row["signal_id"]: (
            sum(len(row[stage]["predicates"]) for stage in ("eligibility", "trigger", "near_miss")),
            len(row["score_components"]),
        )
        for row in contract["signals"]
    }
    for candidate in plan.candidates:
        assert (len(candidate.predicate_results), len(candidate.component_scores)) == expected[candidate.signal_id]
    assert _candidate(plan, "F").state == "insufficient_evidence"


def test_debt_is_per_signal_optional_module_eligibility(tmp_path: Path) -> None:
    manifest = load_workbook_module_manifest()
    profile = {
        "profile_id": "hidden_without_debt",
        "enabled_modules": [
            "core_financial_history",
            "balance_cash_flow",
            "qa_lineage",
            "hidden_value_signals",
        ],
        "profile_pack_ids": [],
        "dimensions": [
            {
                "dimension_id": "total_company",
                "display_name": "Total company",
                "members_source": "universal",
            }
        ],
    }
    manifest["profiles"].append(profile)
    path = tmp_path / "module_manifest.json"
    path.write_text(json.dumps(manifest), encoding="utf-8")

    plan = evaluate_hidden_value_signals(
        _economic_package(),
        profile_id="hidden_without_debt",
        module_manifest_path=path,
    )

    assert len(plan.candidates) == 7
    assert _candidate(plan, "D").state == "unavailable"
    assert "required_modules_disabled:debt_liquidity" in _candidate(plan, "D").reasons
    assert _candidate(plan, "A").triggered is True


def test_conflicting_duplicate_source_candidate_fails_closed() -> None:
    package = _economic_package()
    package["calculation_history"]["quarterly_items"].append(
        _source_row("2025-Q4", "shares_outstanding", 80.0, "m shares")
    )

    candidate = _candidate(evaluate_hidden_value_signals(package, profile_id="anf"), "F")

    assert candidate.state == "invalid_input"
    assert candidate.triggered is False
    assert candidate.score is None


def test_anf_current_package_has_no_unsupported_trigger() -> None:
    aktier_root = next(parent for parent in Path(__file__).resolve().parents if parent.name == "Aktier")
    path = aktier_root / "StockModelData" / "outputs" / "stress_tests" / "ANF_new_ticker_engine" / "ANF_normalized_data_package.json"
    if not path.exists():
        pytest.skip(f"{path} is unavailable")
    package = json.loads(path.read_text(encoding="utf-8"))

    plan = evaluate_hidden_value_signals(package, profile_id="anf", ticker="ANF")

    assert plan.flags_projection == []
    assert plan.state_counts == {"insufficient_evidence": 5, "not_triggered": 1, "unavailable": 1}
    assert _candidate(plan, "A").state == "insufficient_evidence"
    assert _candidate(plan, "C").state == "insufficient_evidence"
    assert _candidate(plan, "D").state == "insufficient_evidence"
    assert _candidate(plan, "F").state == "insufficient_evidence"
    assert _candidate(plan, "G").state == "unavailable"
    assert any(
        "capital-return-unavailable:dividends_paid:TTM through 2026-Q1" in source_ref
        for source_ref in _candidate(plan, "G").source_refs
    )


def test_legacy_cross_ticker_oracles_are_locked_independently() -> None:
    pbi = _legacy_flags("PBI")
    gpre = _legacy_flags("GPRE")
    anf = _legacy_flags("ANF")

    assert pbi["F"] == (100, 1)
    assert all(gpre[signal_id][1] in (0, False) for signal_id in "ABDF")
    assert all(anf[signal_id][1] in (0, False) for signal_id in "ABDF")
