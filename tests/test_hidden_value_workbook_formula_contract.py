from __future__ import annotations

import gc
import json
from pathlib import Path
import shutil
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
import pytest

from pbi_xbrl.hidden_value_signal_economics import evaluate_hidden_value_signals
from pbi_xbrl.hidden_value_workbook_projection import build_hidden_value_workbook_projection
from pbi_xbrl.standard_template_formula_contract import _hidden_value_detail_specs
from pbi_xbrl.standard_template_shell_identity import verify_shell_identity


ROOT = Path(__file__).resolve().parents[1]
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
BINDINGS = ROOT / "docs" / "workbook_binding_map.json"
SIGNALS = ROOT / "docs" / "hidden_value_signal_contract.json"
PERIODS = [f"{year}-Q{quarter}" for year in (2024, 2025) for quarter in range(1, 5)]

BASE_COLUMNS = (
    "metric_key", "metric_id", "value", "unit", "period", "period_role",
    "dimension_id", "member", "status", "reason", "evidence_ids", "source_refs", "formula_ids",
)
AUDIT_COLUMNS = (
    "candidate_key", "signal_id", "display_name", "profile_id", "as_of_period",
    "expected_state", "expected_triggered", "expected_score", "expected_score_denominator",
    "expected_severity", "priority", "reasons", "evidence_ids", "source_refs",
    "resolved_metric_keys", "recompute_record_keys", "module_eligible",
)
RECOMPUTE_COLUMNS = (
    "record_key", "candidate_key", "signal_id", "record_type", "stage", "item_id",
    "metric_key", "operator", "right_metric_key", "threshold", "required_component", "weight",
    "normalization_direction", "normalization_threshold", "normalization_span", "normalization_base",
    "expected_value", "expected_comparison", "expected_passed", "expected_included_weight",
    "expected_normalized_score", "expected_weighted_score", "expected_status",
)
FLAG_COLUMNS = (
    "display_rank", "candidate_key", "signal_id", "display_name", "score", "triggered", "state",
    "severity", "as_of_period", "reason", "evidence_ids", "source_refs", "audit_id",
)


def _source_row(period: str, metric: str, value: float, unit: str) -> dict[str, Any]:
    return {
        "period": period,
        "metric": metric,
        "value": value,
        "unit": unit,
        "status": "source_backed",
        "source_ref": f"fixture:{metric}:{period}",
    }


def _economic_package() -> dict[str, Any]:
    series = {
        "revenue": ([100.0] * 4 + [150.0] * 4, "$m"),
        "operating_income": ([10.0] * 4 + [13.0] * 4, "$m"),
        "base_ebitda": ([15.0] * 4 + [18.75] * 4, "$m"),
        "adjusted_ebitda": ([20.0] * 4 + [37.5] * 4, "$m"),
        "operating_cash_flow": ([20.0] * 4 + [40.0] * 4, "$m"),
        "capital_expenditures": ([5.0] * 8, "$m"),
        "interest_expense": ([4.0] * 4 + [3.0] * 4, "$m"),
        "shares_outstanding": ([100.0] * 4 + [99.5, 99.0, 98.0, 97.0], "m shares"),
        "market_cap": ([700.0] * 4 + [500.0] * 4, "$m"),
        "net_debt": ([300.0] * 4 + [280.0, 250.0, 225.0, 200.0], "$m"),
        "dividend_per_share": ([0.10] * 4 + [0.11, 0.11, 0.12, 0.12], "$/share"),
    }
    return {
        "ticker_metadata": {"ticker": "TEST"},
        "calculation_history": {
            "quarterly_items": [
                _source_row(period, metric, values[index], unit)
                for metric, (values, unit) in series.items()
                for index, period in enumerate(PERIODS)
            ]
        },
        "annual_financials": {"rows": []},
        "valuation_inputs": {
            "price": {
                "value": None,
                "unit": "$/share",
                "period": "2025-Q4",
                "status": "missing_source",
                "source_ref": "fixture:price:missing",
            }
        },
    }


def _write_projection(path: Path, package: dict[str, Any], *, profile_id: str = "anf") -> dict[str, Any]:
    shutil.copyfile(SHELL, path)
    plan = evaluate_hidden_value_signals(package, profile_id=profile_id, ticker="TEST")
    projection = build_hidden_value_workbook_projection(plan).to_dict()
    surfaces = {
        "Hidden_Value_Base": (projection["base_rows"], BASE_COLUMNS),
        "Hidden_Value_Audit": (projection["audit_rows"], AUDIT_COLUMNS),
        "Hidden_Value_Recompute": (projection["recompute_rows"], RECOMPUTE_COLUMNS),
        "Hidden_Value_Flags": (projection["flags_rows"], FLAG_COLUMNS),
    }
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        for sheet_name, (rows, columns) in surfaces.items():
            sheet = workbook[sheet_name]
            for row_number, row in enumerate(rows, start=2):
                for column_number, field_name in enumerate(columns, start=1):
                    sheet.cell(row_number, column_number).value = row.get(field_name)
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(path)
    finally:
        workbook.close()
    return projection


def _formula_cells(workbook: Any, sheet_name: str, targets: tuple[str, ...]) -> list[Any]:
    result = []
    sheet = workbook[sheet_name]
    for target in targets:
        min_col, min_row, max_col, max_row = range_boundaries(target)
        result.extend(
            cell
            for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
            for cell in row
        )
    return result


def test_hidden_value_support_contract_has_exact_headers_formulas_and_protection() -> None:
    workbook = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        assert tuple(workbook["Hidden_Value_Base"].iter_rows(min_row=1, max_row=1, values_only=True))[0][:13] == BASE_COLUMNS
        assert tuple(workbook["Hidden_Value_Audit"].iter_rows(min_row=1, max_row=1, values_only=True))[0][:22] == AUDIT_COLUMNS + (
            "recomputed_state", "recomputed_triggered", "recomputed_score", "parity_status", "discrepancy_reason",
        )
        assert tuple(workbook["Hidden_Value_Recompute"].iter_rows(min_row=1, max_row=1, values_only=True))[0][:50] == RECOMPUTE_COLUMNS + (
            "recomputed_value", "recomputed_aux", "recomputed_result", "recomputed_included_weight",
            "recomputed_normalized_score", "recomputed_weighted_score", "row_parity", "reserved",
            "summary_candidate_key", "summary_signal_id", "eligibility_mode", "trigger_mode", "near_miss_mode",
            "minimum_trigger_predicates", "reweight_available_components", "expected_state", "expected_triggered",
            "expected_score", "expected_score_denominator", "recomputed_eligibility", "recomputed_trigger",
            "recomputed_near_miss", "recomputed_score_denominator", "recomputed_score", "recomputed_state",
            "recomputed_triggered", "candidate_parity",
        )
        assert tuple(workbook["Hidden_Value_Flags"].iter_rows(min_row=1, max_row=1, values_only=True))[0][:13] == FLAG_COLUMNS

        detail = _formula_cells(workbook, "Hidden_Value_Recompute", ("X2:AD92",))
        candidates = _formula_cells(workbook, "Hidden_Value_Recompute", ("AF2:AX8",))
        audit = _formula_cells(workbook, "Hidden_Value_Audit", ("R2:V8",))
        assert (len(detail), len(candidates), len(audit)) == (637, 133, 35)
        assert all(isinstance(cell.value, str) and cell.value.startswith("=") for cell in detail + candidates + audit)
        assert all(cell.protection.locked for cell in detail + candidates + audit)
        assert workbook["Valuation"]["AI139"].value is None
        assert all(workbook[name].sheet_state != "visible" for name in (
            "Hidden_Value_Base", "Hidden_Value_Audit", "Hidden_Value_Recompute", "Hidden_Value_Flags",
        ))
        assert all(workbook[name].protection.sheet for name in (
            "Hidden_Value_Base", "Hidden_Value_Audit", "Hidden_Value_Recompute", "Hidden_Value_Flags",
        ))
    finally:
        workbook.close()


def test_hidden_value_formula_rows_are_compiled_from_the_authoritative_contract() -> None:
    contract = json.loads(SIGNALS.read_text(encoding="utf-8"))
    signals = sorted(contract["signals"], key=lambda row: (row["priority"], row["signal_id"]))
    specs = _hidden_value_detail_specs(signals)

    assert len(specs) == 91
    assert [(row["signal_id"], row["record_type"], row["stage"], row["item_id"]) for row in specs[:3]] == [
        ("A", "required_metric", "required", "ebit_growth_yoy"),
        ("A", "required_metric", "required", "base_ebitda_growth_yoy"),
        ("A", "required_metric", "required", "shares_outstanding_yoy"),
    ]
    assert {row["signal_id"] for row in specs} == set("ABCDEFG")
    source = (ROOT / "pbi_xbrl" / "standard_template_formula_contract.py").read_text(encoding="utf-8")
    hidden_value_source = source[source.index("def _apply_hidden_value_support_formulas"):source.index("def _number_format_for_row")]
    assert "LET(" not in hidden_value_source
    assert "data_only" not in hidden_value_source
    assert "ANF" not in hidden_value_source and "PBI" not in hidden_value_source and "GPRE" not in hidden_value_source


def test_hidden_value_defined_names_are_neutral_exact_and_stale_names_are_absent() -> None:
    workbook = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        expected = {
            "HV_Base_MetricKey": "'Hidden_Value_Base'!$A$2:$A$5001",
            "HV_Base_Value": "'Hidden_Value_Base'!$C$2:$C$5001",
            "HV_Base_Status": "'Hidden_Value_Base'!$I$2:$I$5001",
            "HV_Recompute_CandidateKey": "'Hidden_Value_Recompute'!$AF$2:$AF$8",
            "HV_Recompute_RowParity": "'Hidden_Value_Recompute'!$AD$2:$AD$92",
            "HV_Recompute_CandidateParity": "'Hidden_Value_Recompute'!$AX$2:$AX$8",
            "HV_Flags_CandidateKey": "'Hidden_Value_Flags'!$B$2:$B$8",
            "HV_Flags_Score": "'Hidden_Value_Flags'!$E$2:$E$8",
            "HV_Flags_State": "'Hidden_Value_Flags'!$G$2:$G$8",
        }
        assert {name: workbook.defined_names[name].attr_text for name in expected} == expected
        for stale in ("FCF_TTM_Pos_Years", "Pos_FCF_Ratio", "Interest_Coverage"):
            assert stale not in workbook.defined_names
    finally:
        workbook.close()


def test_formula_mutation_breaks_frozen_shell_identity(tmp_path: Path) -> None:
    drifted = tmp_path / "hidden-value-formula-drift.xlsx"
    shutil.copyfile(SHELL, drifted)
    workbook = load_workbook(drifted, data_only=False, read_only=False)
    try:
        workbook["Hidden_Value_Recompute"]["AU2"] = "=0"
        workbook.save(drifted)
    finally:
        workbook.close()

    report = verify_shell_identity(
        drifted,
        manifest=json.loads(MANIFEST.read_text(encoding="utf-8")),
        binding_payload=json.loads(BINDINGS.read_text(encoding="utf-8")),
    )
    assert report.status == "FAIL"


@pytest.mark.skipif(sys.platform != "win32", reason="Desktop Excel automation is Windows-only")
def test_excel_native_recompute_matches_independent_a_to_g_oracle_and_blocks_mutations(tmp_path: Path) -> None:
    win32com = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")
    path = tmp_path / "hidden-value-excel-parity.xlsx"
    projection = _write_projection(path, _economic_package())
    expected = {"A": 41, "B": 82, "C": 92, "D": 71, "E": 61, "F": 49, "G": 69}
    base_row = 2 + next(
        index for index, row in enumerate(projection["base_rows"]) if row["metric_id"] == "ebit_growth_yoy"
    )
    base_value = projection["base_rows"][base_row - 2]["value"]

    before = load_workbook(path, data_only=True, read_only=False)
    try:
        assert before["Hidden_Value_Recompute"]["AU2"].value is None
    finally:
        before.close()

    excel = None
    book = None
    recompute = None
    audit = None
    base = None
    pythoncom.CoInitialize()
    try:
        excel = win32com.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            book = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=False)
        except Exception as exc:
            pytest.skip(f"Desktop Excel automation could not open the isolated workbook: {exc}")
        excel.CalculateFullRebuild()
        recompute = book.Worksheets("Hidden_Value_Recompute")
        audit = book.Worksheets("Hidden_Value_Audit")
        base = book.Worksheets("Hidden_Value_Base")
        actual = {
            str(recompute.Range(f"AG{row}").Value): int(recompute.Range(f"AU{row}").Value)
            for row in range(2, 9)
        }
        assert actual == expected
        assert all(recompute.Range(f"AV{row}").Value == "triggered" for row in range(2, 9))
        assert all(recompute.Range(f"AX{row}").Value == "PASS" for row in range(2, 9))
        assert all(audit.Range(f"U{row}").Value == "PASS" for row in range(2, 9))

        base.Unprotect()
        audit.Unprotect()
        base.Range(f"C{base_row}").Value = -0.5
        excel.CalculateFullRebuild()
        assert audit.Range("U2").Value == "FAIL"

        base.Range(f"C{base_row}").Value = base_value
        audit.Range("H2").Value = 40
        excel.CalculateFullRebuild()
        assert audit.Range("U2").Value == "FAIL"

        audit.Range("H2").Value = 41
        excel.CalculateFullRebuild()
        assert audit.Range("U2").Value == "PASS"
        base.Protect()
        audit.Protect()
    finally:
        recompute = None
        audit = None
        base = None
        if book is not None:
            book.Close(SaveChanges=False)
            book = None
        if excel is not None:
            excel.Quit()
            excel = None
        gc.collect()
        pythoncom.CoUninitialize()
