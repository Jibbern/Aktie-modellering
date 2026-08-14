import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import pytest
from openpyxl import load_workbook

from tests.workbook_test_resources import delivered_workbook_path

TICKERS = ("PBI", "GPRE", "ANF")
YELLOW_INPUT_FILL = "00FFF2CC"


def _load_workbook(ticker: str, *, data_only: bool = False):
    path = delivered_workbook_path(ticker, Path(__file__).resolve())
    return load_workbook(path, data_only=data_only, read_only=False)


def _ic_sheet(wb: Any, ticker: str):
    sheet = f"{ticker}_Investment_Case"
    assert sheet in wb.sheetnames, f"{ticker}: missing {sheet}"
    return wb[sheet]


def _text(value: Any) -> str:
    return str(value or "").strip()


def _semantic_rgb(value: Any) -> str:
    raw = _text(value).upper()
    assert len(raw) in {6, 8}, f"Unsupported OOXML RGB encoding: {raw!r}"
    return raw[-6:]


def _formula(value: Any) -> str:
    txt = _text(value)
    return txt if txt.startswith("=") else txt


def _compact_formula(value: Any) -> str:
    return re.sub(r"\s+", "", _formula(value)).replace("$", "").upper()


def _row_by_label(ws: Any, label: str, *, start: int = 1, end: int | None = None) -> int:
    end = end or int(ws.max_row or 0)
    for rr in range(start, end + 1):
        if _text(ws.cell(rr, 1).value) == label:
            return rr
    raise AssertionError(f"{ws.title}: could not find row label {label!r}")


def _section_bounds(ws: Any, title: str) -> Tuple[int, int]:
    start = _row_by_label(ws, title)
    end = int(ws.max_row or 0)
    for rr in range(start + 1, int(ws.max_row or 0) + 1):
        first = _text(ws.cell(rr, 1).value)
        fill = _text(ws.cell(rr, 1).fill.fgColor.rgb).upper()
        if first and fill.endswith(("5B9BD5", "4472C4", "6FA8DC")):
            end = rr - 1
            break
    return start, end


def _manual_rows(ws: Any) -> Dict[str, int]:
    start, end = _section_bounds(ws, "Manual Market / Scenario Inputs")
    return {
        _text(ws.cell(rr, 1).value): rr
        for rr in range(start + 1, end + 1)
        if _text(ws.cell(rr, 1).value)
    }


def _bridge_rows(ws: Any) -> Dict[str, int]:
    start, end = _section_bounds(ws, "Scenario Driver Bridge")
    return {
        _text(ws.cell(rr, 1).value): rr
        for rr in range(start + 1, end + 1)
        if _text(ws.cell(rr, 1).value)
    }


def _segment_rows(ws: Any) -> Dict[str, int]:
    try:
        start, end = _section_bounds(ws, "Segment Scenario Inputs")
    except AssertionError:
        return {}
    return {
        _text(ws.cell(rr, 1).value): rr
        for rr in range(start + 1, end + 1)
        if _text(ws.cell(rr, 1).value)
    }


def _tax_rows(wb: Any, ticker: str) -> Dict[str, Dict[str, Any]]:
    assert "Scenario_Bridge_Tax_Treatment" in wb.sheetnames, f"{ticker}: missing tax treatment audit sheet"
    ws = wb["Scenario_Bridge_Tax_Treatment"]
    headers = [_text(ws.cell(1, cc).value) for cc in range(1, int(ws.max_column or 0) + 1)]
    out: Dict[str, Dict[str, Any]] = {}
    for rr in range(2, int(ws.max_row or 0) + 1):
        if _text(ws.cell(rr, 1).value) != ticker:
            continue
        row = {headers[cc - 1]: ws.cell(rr, cc).value for cc in range(1, len(headers) + 1)}
        out[_text(row.get("Bridge item"))] = row
    return out


def _all_formulas(wb: Any, sheet_names: Iterable[str] | None = None) -> List[Tuple[str, str, str]]:
    names = list(sheet_names) if sheet_names is not None else wb.sheetnames
    out: List[Tuple[str, str, str]] = []
    for sheet_name in names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    out.append((sheet_name, cell.coordinate, cell.value))
    return out


def _assert_contains(formula: Any, *parts: str, context: str = "") -> None:
    compact = _compact_formula(formula)
    for part in parts:
        assert part.replace("$", "").upper() in compact, f"{context}: formula {formula!r} missing {part!r}"


def _assert_not_contains(formula: Any, *parts: str, context: str = "") -> None:
    compact = _compact_formula(formula)
    for part in parts:
        assert part.replace("$", "").upper() not in compact, f"{context}: formula {formula!r} unexpectedly contains {part!r}"


def test_manual_inputs_have_yellow_overrides_manual_first_active_formulas_and_tax_rate_at_bottom() -> None:
    required_common = [
        "Forward revenue",
        "Forward EPS",
        "Forward Adj EBITDA",
        "Forward FCF",
        "Operating margin",
        "Diluted shares",
        "Scenario tax rate",
    ]
    ticker_specific = {
        "PBI": ["Cost savings target / run-rate ($m)"],
        "GPRE": ["45Z contribution / guide ($m)", "Crush margin uplift ($m)", "Policy / RVO / E15 / export", "Capex"],
        "ANF": ["Tariff impact (bps)", "Freight tailwind (bps)", "ERP disruption (bps)", "Marketing headwind (bps)", "Buyback amount", "Buyback-adjusted shares"],
    }
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        try:
            ws = _ic_sheet(wb, ticker)
            assert not ws.protection.sheet, f"{ticker}: Investment_Case sheet is protected, manual override cells are not editable"
            rows = _manual_rows(ws)
            for label in required_common + ticker_specific[ticker]:
                rr = rows[label]
                manual = ws.cell(rr, 6)
                active = ws.cell(rr, 7).value
                assert manual.fill.fill_type == "solid", f"{ticker} {label}: manual override cell lacks solid input fill"
                assert _semantic_rgb(manual.fill.fgColor.rgb) == _semantic_rgb(YELLOW_INPUT_FILL), f"{ticker} {label}: manual override fill"
                _assert_contains(active, f"F{rr}<>\"\"", f"F{rr}", context=f"{ticker} {label} active value")
                assert _compact_formula(active).find(f"F{rr}<>\"\"") < _compact_formula(active).find("IF(") + 5 or _compact_formula(active).startswith(f"=IF(F{rr}<>\"\""), (
                    f"{ticker} {label}: manual override should be first branch in active value formula {active!r}"
                )
            scenario_tax_rate_row = rows["Scenario tax rate"]
            next_section = min(
                rr
                for rr in range(scenario_tax_rate_row + 1, int(ws.max_row or 0) + 1)
                if _text(ws.cell(rr, 1).value) in {"Segment Scenario Inputs", "Scenario Driver Bridge"}
            )
            populated_between = [
                _text(ws.cell(rr, 1).value)
                for rr in range(scenario_tax_rate_row + 1, next_section)
                if _text(ws.cell(rr, 1).value)
            ]
            assert not populated_between, f"{ticker}: Scenario tax rate is not the last Manual Inputs row before next section: {populated_between}"
            _assert_contains(ws.cell(scenario_tax_rate_row, 7).value, "0.25", context=f"{ticker} Scenario tax rate active fallback")
        finally:
            wb.close()


def test_valuation_hidden_value_flags_remain_formula_linked_to_audit_sheet() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        try:
            assert "Hidden_Value_Flags" in wb.sheetnames, f"{ticker}: missing Hidden_Value_Flags audit sheet"
            ws = wb["Valuation"]
            flags_header_row = _row_by_label(ws, "Hidden value flags")
            first_visible_row = flags_header_row + 2
            second_visible_row = flags_header_row + 3
            first_helper_formula = _text(ws.cell(first_visible_row, 35).value)
            second_helper_formula = _text(ws.cell(second_visible_row, 35).value)
            label_value = _text(ws.cell(first_visible_row, 1).value)
            title_value = _text(ws.cell(first_visible_row, 2).value)
            score_value = ws.cell(first_visible_row, 6).value
            support_value = _text(ws.cell(first_visible_row, 8).value)
            assert first_helper_formula.replace("'", "") == '=IFERROR(MATCH(1,Hidden_Value_Flags!$L$2:$L$100,0)+1,"")', f"{ticker}: first hidden flag helper"
            audit = wb["Hidden_Value_Flags"]
            active_rows = [
                rr
                for rr in range(2, int(audit.max_row or 0) + 1)
                if _text(audit.cell(rr, 12).value) == "1"
            ]
            if active_rows:
                assert second_helper_formula.replace("'", "") == (
                    f'=IF($AI{first_visible_row}="","",IFERROR(MATCH(1,'
                    f'INDEX(Hidden_Value_Flags!$L:$L,$AI{first_visible_row}+1):'
                    f'Hidden_Value_Flags!$L$100,0)+$AI{first_visible_row},""))'
                ), f"{ticker}: subsequent hidden flag helper"
                first_audit_row = active_rows[0]
                assert label_value == "Flag 1", f"{ticker}: visible flag label identity"
                assert title_value == _text(audit.cell(first_audit_row, 3).value), f"{ticker}: flag summary should stay audit-owned"
                assert score_value == audit.cell(first_audit_row, 4).value, f"{ticker}: flag score should stay audit-owned"
                assert support_value == _text(audit.cell(first_audit_row, 11).value), f"{ticker}: flag support should stay audit-owned"
            else:
                assert label_value == "No triggered flags", f"{ticker}: explicit empty flag state"
                assert title_value == "No scored hidden-value flags currently triggered"
                assert score_value in (None, "")
                assert support_value == "Audit candidates remain in Hidden_Value_Flags / Hidden_Value_Audit."
        finally:
            wb.close()


def test_anf_valuation_guidance_sidepanel_groups_horizons_and_splits_margin_bridge() -> None:
    wb = _load_workbook("ANF", data_only=True)
    try:
        ws = wb["Valuation"]
        header_row = None
        for rr in range(1, 80):
            if _text(ws.cell(rr, 15).value).startswith("Guidance (As of 2026-01-31)"):
                header_row = rr
                break
        assert header_row is not None, "ANF Valuation: missing latest guidance sidepanel"
        rows: List[Tuple[str, str, str]] = []
        for rr in range(header_row + 2, 40):
            metric = _text(ws.cell(rr, 15).value)
            if not metric:
                break
            rows.append((metric, _text(ws.cell(rr, 18).value), _text(ws.cell(rr, 19).value)))
        assert rows, "ANF Valuation: latest guidance sidepanel has no rows"
        seen_q1 = False
        for metric, applies_to, _ in rows:
            if applies_to == "2026-Q1":
                seen_q1 = True
            if seen_q1:
                assert applies_to != "2026 year", f"ANF Valuation: 2026-year row {metric!r} appears after Q1 rows"
        metrics = {metric for metric, _, _ in rows}
        assert "Tariff / margin bridge" not in metrics
        for expected in {"Tariff headwind", "Marketing headwind", "Q1 tariff headwind", "Q1 freight tailwind", "Q1 ERP disruption"}:
            assert expected in metrics, f"ANF Valuation: missing split margin bridge row {expected!r}"
    finally:
        wb.close()


def test_gpre_bridge_driver_formulas_and_tax_treatments_are_incremental_and_type_safe() -> None:
    wb = _load_workbook("GPRE")
    try:
        ws = _ic_sheet(wb, "GPRE")
        bridge = _bridge_rows(ws)
        tax = _tax_rows(wb, "GPRE")

        z45 = bridge["Incremental 45Z uplift vs baseline"]
        assert tax["Incremental 45Z uplift vs baseline"]["Tax treatment"] == "non_taxable_credit"
        assert tax["Incremental 45Z uplift vs baseline"]["Tax rate / conversion used"] == "100% conversion"
        _assert_contains(ws.cell(z45, 2).value, "$C$28", context="GPRE 45Z baseline")
        _assert_contains(ws.cell(z45, 3).value, "$G$28", context="GPRE 45Z active")
        _assert_contains(ws.cell(z45, 4).value, f"C{z45}-B{z45}", context="GPRE 45Z incremental")
        _assert_contains(ws.cell(z45, 5).value, f"D{z45}/$G$16", context="GPRE 45Z EPS")
        _assert_not_contains(ws.cell(z45, 5).value, "$G$31", context="GPRE 45Z should not be taxed")
        _assert_contains(ws.cell(z45, 6).value, f"D{z45}", context="GPRE 45Z EBITDA")
        assert ws.cell(z45, 7).value == 0

        for label in ["Crush margin uplift ($m)", "Policy / RVO / E15 / export"]:
            rr = bridge[label]
            assert tax[label]["Tax treatment"] == "taxable"
            _assert_contains(ws.cell(rr, 5).value, f"D{rr}*(1-$G$31)/$G$16", context=f"GPRE {label} taxable EPS")
            _assert_contains(ws.cell(rr, 6).value, f"D{rr}", context=f"GPRE {label} EBITDA")
            assert ws.cell(rr, 7).value == 0

        capex = bridge["Capex change vs baseline"]
        assert tax["Capex change vs baseline"]["Tax treatment"] == "cash_only"
        _assert_contains(ws.cell(capex, 4).value, f"C{capex}-B{capex}", context="GPRE capex incremental")
        assert ws.cell(capex, 5).value == 0
        assert ws.cell(capex, 6).value == 0
        _assert_contains(ws.cell(capex, 7).value, f"-(C{capex}-B{capex})", context="GPRE capex FCF")

        scenario_formula_blob = "\n".join(
            formula for _, cell, formula in _all_formulas(wb, ["GPRE_Investment_Case"]) if int(re.sub(r"[^0-9]", "", cell) or 0) <= 60
        )
        assert "Economics_Overlay" not in scenario_formula_blob, "GPRE scenario bridge formulas should not depend on Economics_Overlay"
    finally:
        wb.close()


def test_pbi_bridge_cost_savings_interest_capex_and_segment_logic() -> None:
    wb = _load_workbook("PBI")
    try:
        ws = _ic_sheet(wb, "PBI")
        bridge = _bridge_rows(ws)
        tax = _tax_rows(wb, "PBI")
        manual = _manual_rows(ws)

        cost = bridge["Incremental cost savings vs baseline"]
        assert manual["Cost savings target / run-rate ($m)"]
        assert tax["Incremental cost savings vs baseline"]["Tax treatment"] == "taxable"
        _assert_contains(ws.cell(cost, 2).value, "$C$30", context="PBI cost savings baseline")
        _assert_contains(ws.cell(cost, 3).value, "$G$30", context="PBI cost savings active")
        _assert_contains(ws.cell(cost, 4).value, f"C{cost}-B{cost}", context="PBI cost savings incremental")
        _assert_contains(ws.cell(cost, 5).value, f"D{cost}*(1-$G$31)/$G$16", context="PBI cost savings EPS")
        _assert_contains(ws.cell(cost, 6).value, f"D{cost}", context="PBI cost savings EBITDA")

        interest = bridge["Interest/refinancing effect vs baseline"]
        assert tax["Interest/refinancing effect vs baseline"]["Driver type"] == "capital_structure_interest"
        _assert_contains(ws.cell(interest, 4).value, f"B{interest}-C{interest}", context="PBI interest/refi incremental")
        _assert_contains(ws.cell(interest, 5).value, f"D{interest}*(1-$G$31)/$G$16", context="PBI interest/refi EPS")
        assert ws.cell(interest, 6).value == 0
        _assert_contains(ws.cell(interest, 7).value, f"D{interest}", context="PBI interest/refi FCF")
        assert "no interest rate assumed" in _text(ws.cell(bridge["Debt paydown / net debt"], 8).value).lower()

        capex = bridge["Capex change vs baseline"]
        assert tax["Capex change vs baseline"]["Tax treatment"] == "cash_only"
        assert ws.cell(capex, 5).value == 0
        assert ws.cell(capex, 6).value == 0
        _assert_contains(ws.cell(capex, 7).value, f"-(C{capex}-B{capex})", context="PBI capex FCF")

        seg = bridge["Selected segment revenue/margin impact"]
        assert tax["Selected segment revenue/margin impact"]["Tax treatment"] == "taxable"
        _assert_contains(ws.cell(seg, 3).value, 'SUMIF(H35:H36,"Yes",G35:G36)', context="PBI selected segment sum")
        _assert_contains(ws.cell(seg, 5).value, f"D{seg}*(1-$G$31)/$G$16", context="PBI segment EPS")

        segment_rows = _segment_rows(ws)
        for label in ("Presort", "SendTech"):
            rr = segment_rows[label]
            _assert_contains(ws.cell(rr, 5).value, f"C{rr}*D{rr}", context=f"PBI {label} revenue impact")
            _assert_contains(ws.cell(rr, 7).value, f"E{rr}*F{rr}", context=f"PBI {label} EBITDA impact")
            assert "BS_Segments" in _text(ws.cell(rr, 6).value), f"PBI {label}: segment-specific margin should come from BS_Segments before company proxy"
            assert _text(ws.cell(rr, 9).value) == "Segment operating margin"
    finally:
        wb.close()


def test_anf_margin_bridge_buyback_capex_and_segment_active_basis_guardrails() -> None:
    wb = _load_workbook("ANF")
    wb_values = _load_workbook("ANF", data_only=True)
    try:
        ws = _ic_sheet(wb, "ANF")
        ws_values = _ic_sheet(wb_values, "ANF")
        bridge = _bridge_rows(ws)
        tax = _tax_rows(wb, "ANF")

        for label, sign in {
            "Tariff impact (bps)": -1,
            "Freight tailwind (bps)": 1,
            "ERP disruption (bps)": -1,
            "Marketing headwind (bps)": -1,
        }.items():
            rr = _manual_rows(ws)[label]
            source_values = [ws.cell(rr, cc).value for cc in (2, 3, 4, 5, 6)]
            numeric_sources = [value for value in source_values if isinstance(value, (int, float))]
            assert numeric_sources, f"ANF {label}: no source/manual bps value available"
            assert any(value == 0 or (value > 0) == (sign > 0) for value in numeric_sources), (
                f"ANF {label}: source bps sign convention is wrong: {numeric_sources}"
            )
            _assert_contains(
                ws.cell(rr, 7).value,
                f"F{rr}",
                f"C{rr}",
                f"B{rr}",
                f"D{rr}",
                f"E{rr}",
                context=f"ANF {label} active value",
            )

        margin = bridge["Margin bridge vs baseline"]
        assert tax["Margin bridge vs baseline"]["Tax treatment"] == "taxable"
        margin_formula = ws.cell(margin, 3).value
        _assert_contains(margin_formula, "$G$18*SUM", "$G$30", "$G$31", "$G$32", "$G$33", "/10000", context="ANF margin bridge bps to $m")
        _assert_contains(ws.cell(margin, 5).value, f"D{margin}*(1-$G$34)/$G$16", context="ANF margin bridge EPS")
        _assert_contains(ws.cell(margin, 6).value, f"D{margin}", context="ANF margin bridge EBITDA")
        assert ws.cell(margin, 7).value == 0

        buyback = bridge["Buyback/share-count effect"]
        assert tax["Buyback/share-count effect"]["Tax treatment"] == "no_eps_impact"
        _assert_contains(ws.cell(buyback, 5).value, "$G$19*(B51/C51)-$G$19", context="ANF buyback EPS denominator")
        assert ws.cell(buyback, 6).value == 0
        assert ws.cell(buyback, 7).value == 0

        capex = bridge["Capex change vs baseline"]
        assert tax["Capex change vs baseline"]["Tax treatment"] == "cash_only"
        assert ws.cell(capex, 5).value == 0
        assert ws.cell(capex, 6).value == 0
        _assert_contains(ws.cell(capex, 7).value, f"-(C{capex}-B{capex})", context="ANF capex FCF")

        selected_segment = bridge["Selected segment revenue/margin impact"]
        _assert_contains(ws.cell(selected_segment, 3).value, '$B$38="None",0', 'SUMIF(H40:H45,"Yes",G40:G45)', context="ANF selected segment active basis")
        _assert_contains(ws.cell(selected_segment, 5).value, f"D{selected_segment}*(1-$G$34)/$G$16", context="ANF selected segment EPS")

        segment_rows = _segment_rows(ws)
        assert ws.cell(segment_rows["Active basis"], 2).value == "None"
        for label in ("Abercrombie (brand)", "Hollister (brand)"):
            rr = segment_rows[label]
            _assert_contains(ws.cell(rr, 8).value, '$B$38="Brand"', context=f"ANF {label} active basis")
            assert _text(ws.cell(rr, 9).value) == "Active company operating margin proxy"
        for label in ("Americas (geography / stores)", "EMEA (geography / stores)", "APAC (geography / stores)"):
            rr = segment_rows[label]
            _assert_contains(ws.cell(rr, 8).value, '$B$38="Geography"', context=f"ANF {label} active basis")
            assert _text(ws.cell(rr, 9).value) == "Active company operating margin proxy"
    finally:
        wb.close()
        wb_values.close()


def test_gpre_has_no_fake_segment_scenario_inputs() -> None:
    wb = _load_workbook("GPRE")
    try:
        ws = _ic_sheet(wb, "GPRE")
        assert "Segment Scenario Inputs" not in { _text(ws.cell(rr, 1).value) for rr in range(1, int(ws.max_row or 0) + 1) }
        assumptions = wb["Scenario_Driver_Assumptions"]
        rows = [
            [_text(assumptions.cell(rr, cc).value) for cc in range(1, 14)]
            for rr in range(2, int(assumptions.max_row or 0) + 1)
            if _text(assumptions.cell(rr, 1).value) == "GPRE"
        ]
        assert rows == [["GPRE", "Segment Scenario Inputs", "Not enabled", "Disabled", "", "", "", "", "", "", "", "No", "GPRE segment scenario disabled; use ethanol, 45Z, crush and policy drivers."]]
    finally:
        wb.close()


def test_direct_eps_rows_are_not_misclassified_as_non_taxable_credit() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        try:
            tax = _tax_rows(wb, ticker)
            for item, row in tax.items():
                item_low = item.lower()
                if "eps" in item_low and ("adjustment" in item_low or "guide" in item_low or "manual" in item_low):
                    assert row["Tax treatment"] == "direct_eps", f"{ticker} {item}: direct EPS bridge row should use direct_eps"
                assert not (
                    row["Tax treatment"] == "non_taxable_credit"
                    and ("eps" in item_low or "manual" in item_low)
                    and "45z" not in item_low
                ), f"{ticker} {item}: non-45Z direct EPS/manual row misclassified as non_taxable_credit"
        finally:
            wb.close()


def test_bear_base_bull_uses_bridge_adjusted_outputs_and_value_formulas_are_share_scaled() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        try:
            ws = _ic_sheet(wb, ticker)
            bridge = _bridge_rows(ws)
            eps_summary = bridge["Bridge EPS ($/sh)"]
            ebitda_summary = bridge["Bridge Adj EBITDA ($m)"]
            fcf_summary = bridge["Bridge FCF ($m)"]

            scenario_row = _row_by_label(ws, "Scenario")
            for rr in range(scenario_row + 1, scenario_row + 4):
                scenario = _text(ws.cell(rr, 1).value)
                assert scenario in {"Bear", "Base", "Bull"}, f"{ticker}: unexpected scenario row {rr}"
                _assert_contains(ws.cell(rr, 4).value, f"$F${eps_summary}", context=f"{ticker} {scenario} EPS")
                _assert_contains(ws.cell(rr, 5).value, f"$F${ebitda_summary}", context=f"{ticker} {scenario} EBITDA")
                _assert_contains(ws.cell(rr, 6).value, f"$F${fcf_summary}", context=f"{ticker} {scenario} FCF")
                _assert_contains(ws.cell(rr, 7).value, f"D{rr}<=0", '"N/M"', context=f"{ticker} {scenario} P/E guard")
                _assert_contains(ws.cell(rr, 8).value, f"E{rr}*", "-$G$17", "/$G$16", context=f"{ticker} {scenario} EV/EBITDA per share")
                _assert_contains(ws.cell(rr, 9).value, f"F{rr}/", "/$G$16", context=f"{ticker} {scenario} FCF yield per share")

            scenario_matrix_rows = set(range(scenario_row + 1, scenario_row + 4))
            extra_share_price_rows = [
                rr for rr in range(1, int(ws.max_row or 0) + 1)
                if _text(ws.cell(rr, 1).value) in {"Bear", "Base", "Bull"} and _text(ws.cell(rr - 1, 1).value) != "Scenario"
                and rr not in scenario_matrix_rows
            ]
            for rr in extra_share_price_rows:
                row_blob = " ".join(_text(ws.cell(rr, cc).value) for cc in range(1, 5))
                if "IFERROR" in row_blob:
                    assert "/$G$16" in row_blob or "*C" in row_blob, f"{ticker}: value/share row {rr} may not be share-scaled: {row_blob}"
        finally:
            wb.close()


def test_scenario_formulas_do_not_contain_wrong_dependencies_or_double_count_patterns() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        try:
            formulas = _all_formulas(wb)
            bad_undefined = [(sheet, cell, formula) for sheet, cell, formula in formulas if "EBIT_TTM" in formula]
            assert not bad_undefined, f"{ticker}: undefined EBIT_TTM references: {bad_undefined[:3]}"

            ws = _ic_sheet(wb, ticker)
            bridge = _bridge_rows(ws)
            if ticker == "GPRE":
                scenario_formulas = [
                    (cell, formula)
                    for sheet, cell, formula in _all_formulas(wb, ["GPRE_Investment_Case"])
                    if 30 <= int(re.sub(r"[^0-9]", "", cell) or 0) <= 45
                ]
                assert not any("Economics_Overlay" in formula for _, formula in scenario_formulas), "GPRE scenario bridge formula depends on Economics_Overlay"
                z45 = bridge["Incremental 45Z uplift vs baseline"]
                assert "C36-B36" in _compact_formula(ws.cell(z45, 4).value), "GPRE 45Z must use active guide minus baseline"
            for label, rr in bridge.items():
                low = label.lower()
                if "capex" in low:
                    assert ws.cell(rr, 5).value == 0, f"{ticker} {label}: capex should not affect EPS"
                    assert ws.cell(rr, 6).value == 0, f"{ticker} {label}: capex should not affect EBITDA"
                if "buyback" in low or "share-count" in low:
                    assert ws.cell(rr, 6).value == 0, f"{ticker} {label}: buyback/share-count should not affect EBITDA"
                if "segment" in low and ticker == "ANF":
                    _assert_contains(ws.cell(rr, 3).value, '$B$38="None",0', context="ANF segment bridge must honor None basis")
        finally:
            wb.close()
