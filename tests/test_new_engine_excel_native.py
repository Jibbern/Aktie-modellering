from __future__ import annotations

import json
from pathlib import Path
import sys

import pytest
from openpyxl import Workbook, load_workbook

from pbi_xbrl.new_engine_excel import run_excel_native_roundtrip
from pbi_xbrl.new_engine_orchestration import render_shadow, run_plan


ROOT = Path(__file__).resolve().parents[1]
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"


def _package_path() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = (
            parent
            / "StockModelData"
            / "outputs"
            / "stress_tests"
            / "ANF_new_ticker_engine"
            / "ANF_normalized_data_package.json"
        )
        if candidate.exists():
            return candidate
    pytest.fail("ANF normalized package is required for the Excel-native release-path test.")


@pytest.mark.skipif(sys.platform != "win32", reason="Desktop Excel release validation is Windows-only")
def test_real_swedish_excel_roundtrip_uses_owned_process_and_leaves_no_workbook(
    tmp_path: Path,
) -> None:
    package = _package_path()
    plan = run_plan(
        run_dir=tmp_path / "plan",
        package_path=package,
        ticker="ANF",
        profile_id="full_union",
    )
    rendered: dict[str, object] | None = None
    try:
        rendered = render_shadow(
            run_dir=tmp_path / "render",
            output_root=tmp_path / "output",
            version="native-test",
            plan_receipt_path=plan["receipt_path"],
            excel_native="required",
            required_locale_id=1053,
            package_path=package,
            ticker="ANF",
            profile_id="full_union",
        )
        receipt = json.loads(Path(rendered["receipt_path"]).read_text(encoding="utf-8"))
        excel = receipt["validations"]["excel_native"]
        assert excel["status"] == "PASS"
        assert excel["locale_id"] == 1053
        assert excel["formula_error_count"] == 0
        assert excel["owned_process_cleanup"] == "PASS"
        assert isinstance(excel["owned_process_forced_termination"], bool)
        assert excel["macro_part_count"] == 0
        assert excel["external_link_part_count"] == 0
        assert excel["recovery_part_count"] == 0
        assert receipt["validations"]["post_fill"]["status"] == "PASS"
        assert receipt["validations"]["saved_workbook"]["status"] == "PASS"
        formula = receipt["formula_inventory"]
        assert formula["cell_formula_count"] == 2_609
        assert formula["function_counts"]["MAXIFS"] == 324
        assert formula["function_counts"]["MINIFS"] == 324
        assert formula["function_counts"].get("LET", 0) == 0
        assert formula["let_local_occurrences"] == 0
        assert formula["unprefixed_future_functions"] == {}
        assert formula["unsupported_functions"] == {}
    finally:
        if rendered is not None:
            Path(rendered["output_path"]).unlink(missing_ok=True)
            Path(rendered["receipt_path"]).unlink(missing_ok=True)
    assert not list(tmp_path.rglob("*.xlsx"))


@pytest.mark.skipif(sys.platform != "win32", reason="Desktop Excel formula validation is Windows-only")
def test_real_swedish_excel_investment_case_guards_unavailable_and_zero_domains(
    tmp_path: Path,
) -> None:
    source = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        source_sheet = source["{ticker}_Investment_Case"]
        formulas = {
            coordinate: str(source_sheet[coordinate].value)
            for coordinate in (
                "D95",
                "D98",
                "B133",
                "B134",
                "B135",
                "B139",
                "C139",
            )
        }
    finally:
        source.close()

    workbook = Workbook()
    workbook.remove(workbook.active)
    cases: dict[str, dict[str, object]] = {
        "valid_numeric": {},
        "no_adjustment": {"D65": None, "D66": None, "D67": None},
        "unavailable_buyback": {"D65": "Unavailable"},
        "zero_execution_price": {"D66": 0},
        "zero_buyback": {"D65": 0, "D66": "Unavailable"},
        "negative_buyback": {"D65": -1},
        "unavailable_shares": {"F18": "Unavailable"},
        "unavailable_price": {"D106": "Unavailable"},
        "unavailable_wacc": {"D112": "Unavailable"},
        "unavailable_fcf": {"D94": "Unavailable"},
    }
    common: dict[str, object] = {
        "B69": "Base",
        "C83": "Bear",
        "D83": "Base",
        "E83": "Bull",
        "F18": 100,
        "D65": 100,
        "D66": 50,
        "D67": 0,
        "D92": 900,
        "C94": 100,
        "D94": 100,
        "E94": 100,
        "C95": 100,
        "E95": 100,
        "C96": 0,
        "D96": 0,
        "E96": 0,
        "D106": 10,
        "D107": 10,
        "D112": 0.10,
    }
    for title, overrides in cases.items():
        sheet = workbook.create_sheet(title)
        for coordinate, value in (common | overrides).items():
            sheet[coordinate] = value
        for coordinate, formula in formulas.items():
            sheet[coordinate] = formula

    path = tmp_path / "investment_case_formula_guards.xlsx"
    workbook.save(path)
    workbook.close()

    result = run_excel_native_roundtrip(
        path,
        ticker="ANF",
        required_locale_id=1053,
    )

    assert result["status"] == "PASS"
    assert result["locale_id"] == 1053
    assert result["recalculation_count"] == 2
    assert result["formula_error_count"] == 0
    assert result["macro_part_count"] == 0
    assert result["external_link_part_count"] == 0
    assert result["recovery_part_count"] == 0
    assert result["owned_process_cleanup"] == "PASS"

    calculated = load_workbook(path, data_only=True, read_only=False)
    try:
        valid = calculated["valid_numeric"]
        assert valid["D95"].value == pytest.approx(98)
        assert valid["D98"].value == pytest.approx(900 / 98)
        assert valid["B133"].value == pytest.approx(980)
        assert valid["B135"].value == pytest.approx(980)
        assert valid["B139"].value == pytest.approx((980 * 0.10 - 100) / 1080)
        assert valid["C139"].value == "Market EV, selected FCF and WACC"

        no_adjustment = calculated["no_adjustment"]
        assert no_adjustment["D95"].value == pytest.approx(100)
        assert no_adjustment["D98"].value == pytest.approx(9)

        zero_buyback = calculated["zero_buyback"]
        assert zero_buyback["D95"].value == pytest.approx(100)
        assert zero_buyback["D98"].value == pytest.approx(9)

        for title in (
            "unavailable_buyback",
            "zero_execution_price",
            "negative_buyback",
            "unavailable_shares",
        ):
            sheet = calculated[title]
            assert sheet["D95"].value in (None, "")
            assert sheet["D98"].value in (None, "")
        assert calculated["unavailable_price"]["B133"].value in (None, "")

        for title in (
            "unavailable_wacc",
            "unavailable_fcf",
        ):
            sheet = calculated[title]
            assert sheet["B139"].value in (None, "")
            assert sheet["C139"].value == "Unavailable | Market EV, selected FCF and WACC"
    finally:
        calculated.close()
