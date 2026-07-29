from __future__ import annotations

from copy import deepcopy
import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

import pbi_xbrl.new_engine_orchestration as orchestration

from pbi_xbrl.new_engine_orchestration import (
    NewEngineOrchestrationError,
    _saved_workbook_validation,
    _verify_formula_inventory_semantics,
    render_shadow,
    run_plan,
    validate_workbook_immutable,
)
from pbi_xbrl.new_ticker_value_filler import fill_standard_template_from_package
from pbi_xbrl.workbook_validation_runner import (
    ValidationConfig,
    ValidationIssue,
    WorkbookValidationResult,
    validate_workbook,
)


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "standard_stock_model_template.xlsx"
MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
MODULE_MANIFEST = ROOT / "docs" / "workbook_module_manifest.json"
STYLE_POLICY = ROOT / "docs" / "standard_template_style_policy.json"


def _package_path() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = (
            parent
            / "StockModelData"
            / "outputs"
            / "stress_tests"
            / "ANF_new_ticker_engine"
            / "ANF_normalized_data_package.json"
        )
        if candidate.exists():
            return candidate
    pytest.skip("ANF normalized package is unavailable for full orchestration testing.")


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _kwargs() -> dict[str, object]:
    return {
        "package_path": _package_path(),
        "ticker": "ANF",
        "profile_id": "full_union",
        "template_path": TEMPLATE,
        "manifest_path": MANIFEST,
        "binding_map_path": BINDING_MAP,
        "module_manifest_path": MODULE_MANIFEST,
        "style_policy_path": STYLE_POLICY,
    }


@pytest.fixture(scope="module")
def anf_plan(tmp_path_factory: pytest.TempPathFactory) -> dict[str, object]:
    return run_plan(run_dir=tmp_path_factory.mktemp("new-engine-plan") / "plan", **_kwargs())


def test_plan_writes_reproducible_plans_and_non_authoritative_receipt(anf_plan: dict[str, object]) -> None:
    result = anf_plan

    assert result["status"] == "PASS"
    assert result["binding_plan_path"].name == "binding_plan.json"
    assert result["style_plan_path"].name == "style_plan.json"
    receipt = json.loads(result["receipt_path"].read_text(encoding="utf-8"))
    assert receipt["receipt_version"] == "new-engine-run/v1"
    assert receipt["command"] == "plan"
    assert receipt["contract_profile"]["profile_id"] == "full_union"
    assert receipt["plans"]["binding"]["digest"]
    assert receipt["plans"]["binding"]["planned_write_count"] == 23_521
    assert receipt["plans"]["binding"]["structured_skip_count"] == 2_012
    assert receipt["plans"]["binding"]["issue_count"] == 761
    assert receipt["plans"]["binding"]["occurrence_count"] == 2_323
    assert receipt["plans"]["binding"]["blocking_issue_count"] == 0
    assert receipt["plans"]["style"]["digest"]
    assert receipt["plans"]["style"]["action_count"] == 770
    assert receipt["plans"]["style"]["decision_count"] == 1_298
    assert receipt["formula_inventory"]["cell_formula_count"] == 2_690


def test_ticker_profile_and_digest_mismatch_fail_before_artifacts(tmp_path: Path) -> None:
    for overrides, match in (
        ({"ticker": "WRONG"}, "ticker"),
        ({"profile_id": "anf"}, "profile"),
        ({"expected_contract_digest": "0" * 64}, "contract"),
        ({"expected_binding_plan_digest": "0" * 64}, "binding/value plan"),
        ({"expected_style_plan_digest": "0" * 64}, "style plan"),
    ):
        kwargs = _kwargs()
        kwargs.update(overrides)
        run_dir = tmp_path / match
        with pytest.raises(NewEngineOrchestrationError, match=match):
            run_plan(run_dir=run_dir, **kwargs)
        assert not run_dir.exists()


def test_forged_plan_receipt_cannot_authorize_blocked_package(
    tmp_path: Path, anf_plan: dict[str, object]
) -> None:
    valid = anf_plan
    package = json.loads(_package_path().read_text(encoding="utf-8"))
    package["company_profile"].pop("revenue_streams")
    blocked_path = tmp_path / "blocked.json"
    blocked_path.write_text(json.dumps(package), encoding="utf-8")
    kwargs = _kwargs()
    kwargs["package_path"] = blocked_path

    with pytest.raises(NewEngineOrchestrationError, match="blocker"):
        render_shadow(
            run_dir=tmp_path / "blocked-run",
            output_root=tmp_path / "output",
            version="v1",
            plan_receipt_path=valid["receipt_path"],
            excel_native="off",
            **kwargs,
        )
    assert not list((tmp_path / "output").glob("*.xlsx")) if (tmp_path / "output").exists() else True


def test_render_shadow_is_no_overwrite_and_cleans_failed_candidate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, anf_plan: dict[str, object]
) -> None:
    plan = anf_plan
    output_root = tmp_path / "output"
    monkeypatch.setattr(
        "pbi_xbrl.new_engine_orchestration._saved_workbook_validation",
        lambda _path, _ticker, **_kwargs: {"status": "PASS", "overall": "PASS", "issues": []},
    )

    rendered = render_shadow(
        run_dir=tmp_path / "render-run",
        output_root=output_root,
        version="v1",
        plan_receipt_path=plan["receipt_path"],
        excel_native="off",
        **_kwargs(),
    )

    assert rendered["status"] == "PASS"
    assert rendered["output_path"].name == "ANF_shadow_model_v1.xlsx"
    assert rendered["receipt_path"].name == "ANF_shadow_model_v1.run.json"
    before = rendered["output_path"].read_bytes()
    with pytest.raises(NewEngineOrchestrationError, match="already exists"):
        render_shadow(
            run_dir=tmp_path / "second-run",
            output_root=output_root,
            version="v1",
            plan_receipt_path=plan["receipt_path"],
            excel_native="off",
            **_kwargs(),
        )
    assert rendered["output_path"].read_bytes() == before
    assert not list(output_root.glob("*.candidate.xlsx"))


def test_validate_is_immutable_even_when_excel_uses_an_isolated_copy(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, anf_plan: dict[str, object]
) -> None:
    workbook = tmp_path / "input.xlsx"
    workbook.write_bytes(TEMPLATE.read_bytes())
    before = _sha256(workbook)
    plan = anf_plan
    observed: list[Path] = []
    observed_native_modes: list[bool] = []
    observed_saved_binding_maps: list[tuple[Path, str]] = []

    def strict_post_fill(*_args: object, **kwargs: object) -> dict[str, object]:
        observed_native_modes.append(bool(kwargs.get("excel_native_roundtrip")))
        return {"status": "PASS", "issues": []}

    monkeypatch.setattr(
        "pbi_xbrl.new_engine_orchestration._strict_post_fill_validation",
        strict_post_fill,
    )
    def saved_validation(_path: Path, _ticker: str) -> dict[str, object]:
        selected = orchestration._SAVED_VALIDATION_BINDING_MAP.get()
        assert selected is not None
        observed_saved_binding_maps.append(selected)
        return {"status": "PASS", "overall": "PASS", "issues": []}

    monkeypatch.setattr(
        "pbi_xbrl.new_engine_orchestration._saved_workbook_validation",
        saved_validation,
    )

    def excel_roundtrip(path: Path, **_kwargs: object) -> dict[str, object]:
        observed.append(path)
        path.write_bytes(path.read_bytes() + b"excel-save")
        return {"status": "PASS", "formula_error_count": 0}

    monkeypatch.setattr("pbi_xbrl.new_engine_orchestration.run_excel_native_roundtrip", excel_roundtrip)

    result = validate_workbook_immutable(
        workbook_path=workbook,
        run_dir=tmp_path / "validate-run",
        plan_receipt_path=plan["receipt_path"],
        excel_native="required",
        **_kwargs(),
    )

    assert result["status"] == "PASS"
    assert _sha256(workbook) == before
    assert observed and observed[0] != workbook
    assert not observed[0].exists()
    assert observed_native_modes == [True, True]
    assert observed_saved_binding_maps == [(BINDING_MAP.resolve(), _sha256(BINDING_MAP))] * 2


def test_validate_accepts_current_filled_shadow_without_mutating_input(
    tmp_path: Path, anf_plan: dict[str, object]
) -> None:
    workbook = tmp_path / "ANF_shadow_model_current.xlsx"
    fill_standard_template_from_package(_package_path(), output_path=workbook)
    before = _sha256(workbook)

    result = validate_workbook_immutable(
        workbook_path=workbook,
        run_dir=tmp_path / "validate-excel-saved",
        plan_receipt_path=anf_plan["receipt_path"],
        excel_native="off",
        **_kwargs(),
    )

    assert result["status"] == "PASS"
    assert result["receipt"]["output"]["immutable_input"] is True
    assert _sha256(workbook) == before


def test_validate_failure_preserves_preexisting_unowned_receipt(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, anf_plan: dict[str, object]
) -> None:
    workbook = tmp_path / "input.xlsx"
    workbook.write_bytes(TEMPLATE.read_bytes())
    run_dir = tmp_path / "validate-run"
    run_dir.mkdir()
    receipt_path = run_dir / "run_receipt.json"
    original_receipt = b'{"owner":"another-run"}\n'
    receipt_path.write_bytes(original_receipt)

    def fail_post_fill(*_args: object, **_kwargs: object) -> dict[str, object]:
        raise NewEngineOrchestrationError("injected immutable validation failure")

    monkeypatch.setattr(
        "pbi_xbrl.new_engine_orchestration._strict_post_fill_validation",
        fail_post_fill,
    )

    with pytest.raises(NewEngineOrchestrationError, match="injected immutable"):
        validate_workbook_immutable(
            workbook_path=workbook,
            run_dir=run_dir,
            plan_receipt_path=anf_plan["receipt_path"],
            excel_native="off",
            **_kwargs(),
        )

    assert receipt_path.read_bytes() == original_receipt
    assert not (run_dir / "binding_plan.json").exists()
    assert not (run_dir / "style_plan.json").exists()


def test_tampered_receipt_cannot_authorize_render(
    tmp_path: Path, anf_plan: dict[str, object]
) -> None:
    receipt = json.loads(Path(anf_plan["receipt_path"]).read_text(encoding="utf-8"))
    receipt["plans"]["binding"]["digest"] = "f" * 64
    tampered = tmp_path / "tampered.json"
    tampered.write_text(json.dumps(receipt), encoding="utf-8")

    with pytest.raises(NewEngineOrchestrationError, match="tampered"):
        render_shadow(
            run_dir=tmp_path / "run",
            output_root=tmp_path / "output",
            version="v1",
            plan_receipt_path=tampered,
            excel_native="off",
            **_kwargs(),
        )
    assert not (tmp_path / "run").exists()
    assert not (tmp_path / "output").exists()


def test_render_validation_failure_removes_only_owned_candidate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, anf_plan: dict[str, object]
) -> None:
    output_root = tmp_path / "output"

    def fail_post_fill(*_args: object, **_kwargs: object) -> dict[str, object]:
        raise NewEngineOrchestrationError("injected post-fill validation failure")

    monkeypatch.setattr(
        "pbi_xbrl.new_engine_orchestration._strict_post_fill_validation",
        fail_post_fill,
    )
    with pytest.raises(NewEngineOrchestrationError, match="injected"):
        render_shadow(
            run_dir=tmp_path / "run",
            output_root=output_root,
            version="v2",
            plan_receipt_path=anf_plan["receipt_path"],
            excel_native="off",
            **_kwargs(),
        )

    assert not (output_root / "ANF_shadow_model_v2.xlsx").exists()
    assert not (output_root / "ANF_shadow_model_v2.run.json").exists()
    assert not list(output_root.glob("*.candidate.xlsx"))


def test_publication_race_preserves_racer_and_removes_candidate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, anf_plan: dict[str, object]
) -> None:
    import pbi_xbrl.new_engine_orchestration as orchestration

    output_root = tmp_path / "output"
    original_publish = orchestration.publish_no_overwrite

    monkeypatch.setattr(
        orchestration,
        "_saved_workbook_validation",
        lambda _path, _ticker, **_kwargs: {"status": "PASS", "overall": "PASS", "issues": []},
    )

    def race(candidate: Path, final: Path, **kwargs: object) -> None:
        if final.suffix == ".xlsx":
            final.write_bytes(b"racing publisher")
        original_publish(candidate, final, **kwargs)

    monkeypatch.setattr(orchestration, "publish_no_overwrite", race)
    with pytest.raises(NewEngineOrchestrationError, match="already exists"):
        render_shadow(
            run_dir=tmp_path / "run",
            output_root=output_root,
            version="v3",
            plan_receipt_path=anf_plan["receipt_path"],
            excel_native="off",
            **_kwargs(),
        )

    assert (output_root / "ANF_shadow_model_v3.xlsx").read_bytes() == b"racing publisher"
    assert not list(output_root.glob("*.candidate.xlsx"))


def test_saved_gate_reports_prose_quarter_labels_as_advisory_but_keeps_blockers(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    workbook = tmp_path / "workbook.xlsx"
    workbook.write_bytes(b"fixture")
    advisory = WorkbookValidationResult(
        ticker="ANF",
        path=str(workbook),
        quarter_label_advisory_count=1,
        issues=[
            ValidationIssue(
                category="quarter_label",
                sheet="Quarter_Notes_UI",
                cell="M10",
                value="Q4 2025 earnings call",
                classification="advisory",
            )
        ],
    )
    monkeypatch.setattr(
        "pbi_xbrl.new_engine_orchestration.validate_workbook",
        lambda *_args, **_kwargs: advisory,
    )

    report = _saved_workbook_validation(workbook, "ANF")

    assert report["status"] == "PASS"
    assert report["runner_overall"] == "PASS"
    assert [row["category"] for row in report["advisory_issues"]] == ["quarter_label"]

    blocking = WorkbookValidationResult(
        ticker="ANF",
        path=str(workbook),
        missing_required_sheets=["Valuation"],
    )
    monkeypatch.setattr(
        "pbi_xbrl.new_engine_orchestration.validate_workbook",
        lambda *_args, **_kwargs: blocking,
    )
    with pytest.raises(NewEngineOrchestrationError, match="Saved-workbook validation failed"):
        _saved_workbook_validation(workbook, "ANF")


def test_nondefault_binding_map_is_shared_by_plan_saved_and_excel_validation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    selected_map = tmp_path / "selected_binding_map.json"
    selected_map.write_bytes(BINDING_MAP.read_bytes())
    selected_hash = _sha256(selected_map)
    kwargs = _kwargs()
    kwargs["binding_map_path"] = selected_map

    planned = run_plan(run_dir=tmp_path / "plan", **kwargs)
    receipt = json.loads(Path(planned["receipt_path"]).read_text(encoding="utf-8"))
    assert receipt["inputs"]["binding_map"] == {
        "path": str(selected_map.resolve()),
        "size": selected_map.stat().st_size,
        "sha256": selected_hash,
    }

    workbook = tmp_path / "ANF_selected_binding_map.xlsx"
    fill_standard_template_from_package(
        _package_path(),
        output_path=workbook,
        ticker_override="ANF",
        template_path=TEMPLATE,
        manifest_path=MANIFEST,
        binding_map_path=selected_map,
        module_manifest_path=MODULE_MANIFEST,
        style_policy_path=STYLE_POLICY,
    )

    observed_paths: list[Path] = []
    original_validate = orchestration.validate_workbook

    def capture_validate(
        path: Path,
        ticker: str,
        *,
        config: ValidationConfig | None = None,
    ) -> WorkbookValidationResult:
        assert config is not None
        observed_paths.append(Path(config.binding_map_path).resolve())
        return original_validate(path, ticker, config=config)

    monkeypatch.setattr(orchestration, "validate_workbook", capture_validate)
    with orchestration._binding_map_validation_scope(selected_map, selected_hash):
        report = _saved_workbook_validation(workbook, "ANF")
    assert report["status"] == "PASS"
    assert report["binding_map_identity"]["path"] == str(selected_map.resolve())
    assert report["binding_map_identity"]["sha256"] == selected_hash
    assert report["quarter_label_advisories"] == 6
    assert observed_paths == [selected_map.resolve()]

    wb = load_workbook(workbook)
    wb["Operating_Drivers"]["H13"] = "Use Q4 2025 as the current baseline."
    wb.save(workbook)
    wb.close()
    with pytest.raises(NewEngineOrchestrationError, match="Saved-workbook validation failed"):
        with orchestration._binding_map_validation_scope(selected_map, selected_hash):
            _saved_workbook_validation(workbook, "ANF")
    wb = load_workbook(workbook)
    wb["Operating_Drivers"]["H13"] = (
        "Use FY2025 year-end results as the baseline for sales, margin and earnings momentum."
    )
    wb.save(workbook)
    wb.close()

    payload = json.loads(selected_map.read_text(encoding="utf-8"))
    quarter_binding = next(
        row for row in payload["bindings"] if row["binding_id"] == "qn_quarter_note_rows"
    )
    for collection_name in ("row_schema", "target_columns"):
        for column in quarter_binding[collection_name]:
            if column.get("source_field") == "source_display":
                column["target_role"] = "current_claim"
    selected_map.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    changed_hash = _sha256(selected_map)

    with pytest.raises(NewEngineOrchestrationError, match="changed after planning"):
        with orchestration._binding_map_validation_scope(selected_map, selected_hash):
            _saved_workbook_validation(workbook, "ANF")

    changed_result = validate_workbook(
        workbook,
        "ANF",
        config=ValidationConfig(
            enable_quality_guardrails=False,
            binding_map_path=selected_map,
        ),
    )
    assert changed_result.quarter_label_issue_count == 6
    assert changed_result.quarter_label_advisory_count == 0
    assert all(
        issue.classification == "blocking" and "current_claim" in issue.detail
        for issue in changed_result.issues
        if issue.category == "quarter_label"
    )
    with pytest.raises(NewEngineOrchestrationError, match="Saved-workbook validation failed"):
        with orchestration._binding_map_validation_scope(selected_map, changed_hash):
            _saved_workbook_validation(workbook, "ANF")

    missing_map = tmp_path / "missing_binding_map.json"
    missing_map.write_bytes(BINDING_MAP.read_bytes())
    missing_hash = _sha256(missing_map)
    missing_map.unlink()
    with pytest.raises(NewEngineOrchestrationError, match="Required input does not exist"):
        with orchestration._binding_map_validation_scope(missing_map, missing_hash):
            _saved_workbook_validation(workbook, "ANF")


def test_excel_shared_formula_inventory_may_reencode_but_semantics_must_match() -> None:
    expected = {
        "cell_formula_count": 2_141,
        "function_counts": {"MAXIFS": 324, "MINIFS": 324, "LET": 4, "IF": 3_673},
        "let_local_occurrences": 204,
        "future_function_cell_count": 232,
    }
    excel_saved = deepcopy(expected)
    excel_saved["function_counts"]["IF"] = 2_593

    assert _verify_formula_inventory_semantics(expected, excel_saved)["status"] == "PASS"

    excel_saved["function_counts"]["MAXIFS"] = 323
    with pytest.raises(NewEngineOrchestrationError, match="semantically"):
        _verify_formula_inventory_semantics(expected, excel_saved)
