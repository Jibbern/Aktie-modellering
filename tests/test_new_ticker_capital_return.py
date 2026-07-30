from __future__ import annotations

from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook
import pytest

from pbi_xbrl.new_ticker_capital_return import (
    CAPITAL_RETURN_PRODUCT_ROWS,
    CapitalReturnResolutionError,
    build_capital_return_workbook_projection,
    derive_cash_per_program_share,
    derive_fcf_coverage,
    derive_net_share_reduction,
    derive_total_capital_return,
    make_capital_return_record,
    make_unavailable_record,
    validate_capital_return_records,
)
from scripts.materialize_standard_template_shell import (
    _configure_valuation_capital_return_product,
)


SHA = "a" * 64


def _record(
    metric_id: str,
    value: float,
    *,
    unit: str,
    currency: str,
    semantic_role: str | None = None,
    fiscal_period: str = "2026-Q1",
    period_type: str = "quarter",
    period_start: str = "2026-02-01",
    period_end: str = "2026-05-02",
    duration_or_instant: str = "duration",
    aggregation_role: str = "additive_flow",
) -> dict:
    return make_capital_return_record(
        metric_id=metric_id,
        semantic_role=semantic_role or metric_id,
        fiscal_period=fiscal_period,
        period_type=period_type,
        period_start=period_start,
        period_end=period_end,
        duration_or_instant=duration_or_instant,
        publication_date="2026-06-05",
        source_document="sec_cache/TEST/filing.htm",
        source_document_sha256=SHA,
        source_section="test source",
        unit=unit,
        currency=currency,
        scale="millions" if unit in {"$m", "m shares"} else "ratio",
        source_classification="source_native_numeric",
        aggregation_role=aggregation_role,
        evidence_ref=f"sec_cache/TEST/filing.htm#{metric_id}:{fiscal_period}",
        value=value,
        source_alias="TEST filing",
    )


def _unavailable(
    metric_id: str,
    *,
    fiscal_period: str,
    period_type: str,
    period_start: str,
    period_end: str,
    unit: str,
) -> dict:
    return make_unavailable_record(
        metric_id=metric_id,
        semantic_role=metric_id,
        fiscal_period=fiscal_period,
        period_type=period_type,
        period_start=period_start,
        period_end=period_end,
        duration_or_instant="duration",
        publication_date="2026-06-05",
        source_document="sec_cache/TEST/filing.htm",
        source_document_sha256=SHA,
        source_section="bounded evidence review",
        unit=unit,
        currency="USD" if unit == "$m" else "not_applicable",
        scale="millions",
        aggregation_role="additive_flow",
        evidence_ref=f"sec_cache/TEST/filing.htm#{metric_id}:{fiscal_period}",
        reason="No accepted source-native value.",
        source_alias="TEST filing",
    )


def test_cash_per_program_share_requires_exact_program_identity_and_period() -> None:
    cash = _record("repurchase_cash_program", 105.018, unit="$m", currency="USD")
    shares = _record(
        "accounting_program_shares_repurchased",
        1.156,
        unit="m shares",
        currency="not_applicable",
    )
    derived = derive_cash_per_program_share(cash, shares)
    assert derived["value"] == pytest.approx(90.846021)
    assert derived["source_classification"] == "derived_exact"
    assert derived["metric_id"] == "cash_per_program_share"

    withholding = deepcopy(shares)
    withholding["metric_id"] = "employee_tax_withholding_shares"
    with pytest.raises(CapitalReturnResolutionError, match="wrong identity"):
        derive_cash_per_program_share(cash, withholding)

    wrong_period = deepcopy(shares)
    wrong_period["fiscal_period"] = "2025-Q4"
    with pytest.raises(CapitalReturnResolutionError, match="period-incompatible"):
        derive_cash_per_program_share(cash, wrong_period)

    zero_shares = deepcopy(shares)
    zero_shares["value"] = 0
    with pytest.raises(CapitalReturnResolutionError, match="must be positive"):
        derive_cash_per_program_share(cash, zero_shares)


def test_net_share_reduction_requires_exact_roll_forward_snapshots() -> None:
    repurchases = _record(
        "accounting_program_shares_repurchased",
        1.156,
        unit="m shares",
        currency="not_applicable",
    )
    issuance = _record(
        "share_issuance_sbc",
        0.582,
        unit="m shares",
        currency="not_applicable",
    )
    beginning = _record(
        "beginning_period_end_shares",
        45.005,
        unit="m shares",
        currency="not_applicable",
        period_type="point_in_time",
        period_start="2026-01-31",
        period_end="2026-01-31",
        duration_or_instant="instant",
        aggregation_role="point_in_time",
    )
    ending = _record(
        "ending_period_end_shares",
        44.431,
        unit="m shares",
        currency="not_applicable",
        period_type="point_in_time",
        period_start="2026-05-02",
        period_end="2026-05-02",
        duration_or_instant="instant",
        aggregation_role="point_in_time",
    )
    derived = derive_net_share_reduction(
        repurchases,
        issuance,
        beginning_shares_record=beginning,
        ending_shares_record=ending,
    )
    assert derived["value"] == pytest.approx(0.574)
    assert set(derived["component_record_ids"]) == {
        repurchases["record_id"],
        issuance["record_id"],
        beginning["record_id"],
        ending["record_id"],
    }

    weighted_average = deepcopy(beginning)
    weighted_average["metric_id"] = "diluted_weighted_average_shares"
    with pytest.raises(CapitalReturnResolutionError, match="beginning identity"):
        derive_net_share_reduction(
            repurchases,
            issuance,
            beginning_shares_record=weighted_average,
            ending_shares_record=ending,
        )

    incompatible_end = deepcopy(ending)
    incompatible_end["value"] = 44.5
    with pytest.raises(CapitalReturnResolutionError, match="do not reconcile"):
        derive_net_share_reduction(
            repurchases,
            issuance,
            beginning_shares_record=beginning,
            ending_shares_record=incompatible_end,
        )


def test_total_return_and_fcf_coverage_fail_closed() -> None:
    buybacks = _record("repurchase_cash_program", 100, unit="$m", currency="USD")
    dividends = _record("dividends_paid", 20, unit="$m", currency="USD")
    total = derive_total_capital_return(buybacks, dividends)
    assert total["value"] == 120

    fcf = _record("free_cash_flow", 200, unit="$m", currency="USD")
    coverage = derive_fcf_coverage(buybacks, fcf, metric_id="buybacks_to_fcf")
    assert coverage["value"] == 0.5
    for invalid_fcf in (0, -1):
        bad = deepcopy(fcf)
        bad["value"] = invalid_fcf
        with pytest.raises(CapitalReturnResolutionError, match="zero or negative"):
            derive_fcf_coverage(buybacks, bad, metric_id="buybacks_to_fcf")

    wrong_period = deepcopy(fcf)
    wrong_period["fiscal_period"] = "2025-Q4"
    with pytest.raises(CapitalReturnResolutionError, match="period-incompatible"):
        derive_fcf_coverage(buybacks, wrong_period, metric_id="buybacks_to_fcf")


def test_projection_preserves_numeric_zero_unavailable_and_source_order_independence() -> None:
    records = [
        _record("dividends_paid", 0, unit="$m", currency="USD"),
        _unavailable(
            "dividends_paid",
            fiscal_period="TTM through 2026-Q1",
            period_type="ttm",
            period_start="2025-05-04",
            period_end="2026-05-02",
            unit="$m",
        ),
        _unavailable(
            "dividends_paid",
            fiscal_period="2025-FY",
            period_type="annual",
            period_start="2025-02-02",
            period_end="2026-01-31",
            unit="$m",
        ),
    ]
    package = {"capital_returns": {"collection_version": "1.0.0", "records": records}}
    forward = build_capital_return_workbook_projection(package)
    reverse_package = deepcopy(package)
    reverse_package["capital_returns"]["records"].reverse()
    reverse = build_capital_return_workbook_projection(reverse_package)

    row = next(row for row in forward.product_rows if row["row_key"] == "dividends_paid")
    assert row["latest_quarter"] == 0
    assert row["ttm"] is None
    assert row["latest_completed_year"] is None
    assert forward.projection_digest == reverse.projection_digest
    assert validate_capital_return_records(records) == validate_capital_return_records(
        reversed(records)
    )


def test_profiles_without_typed_collection_get_coherent_unavailable_product() -> None:
    package = {
        "capital_returns": {"buybacks": {"status": "missing_source", "value": None}},
        "quarterly_financials": {"rows": [{"period": "2026-Q1"}]},
        "annual_financials": {"rows": [{"period": "2025-FY"}]},
    }
    projection = build_capital_return_workbook_projection(package)
    assert projection.collection_state == "unavailable"
    assert projection.latest_quarter_label == "2026-Q1"
    assert projection.ttm_label == "TTM through 2026-Q1"
    assert projection.annual_label == "2025-FY"
    assert len(projection.product_rows) == len(CAPITAL_RETURN_PRODUCT_ROWS) == 15
    assert all(row["state_context"].startswith("Unavailable:") for row in projection.product_rows)

    with pytest.raises(CapitalReturnResolutionError, match="no records collection"):
        build_capital_return_workbook_projection(
            {"capital_returns": {"collection_version": "1.0.0"}}
        )


def test_buyback_dividend_and_mixed_profiles_remain_distinct_and_coherent() -> None:
    periods = (
        ("2026-Q1", "quarter", "2026-02-01", "2026-05-02"),
        ("TTM through 2026-Q1", "ttm", "2025-05-04", "2026-05-02"),
        ("2025-FY", "annual", "2025-02-02", "2026-01-31"),
    )

    def records_for(metric_id: str, value: float) -> list[dict]:
        return [
            _record(
                metric_id,
                value,
                unit="$m",
                currency="USD",
                fiscal_period=fiscal_period,
                period_type=period_type,
                period_start=period_start,
                period_end=period_end,
            )
            for fiscal_period, period_type, period_start, period_end in periods
        ]

    def unavailable_for(metric_id: str) -> list[dict]:
        return [
            _unavailable(
                metric_id,
                fiscal_period=fiscal_period,
                period_type=period_type,
                period_start=period_start,
                period_end=period_end,
                unit="$m",
            )
            for fiscal_period, period_type, period_start, period_end in periods
        ]

    buyback_only = build_capital_return_workbook_projection(
        {
            "capital_returns": {
                "collection_version": "1.0.0",
                "records": [
                    *records_for("repurchase_cash_program", 25),
                    *unavailable_for("dividends_paid"),
                ],
            }
        }
    )
    dividend_only = build_capital_return_workbook_projection(
        {
            "capital_returns": {
                "collection_version": "1.0.0",
                "records": [
                    *unavailable_for("repurchase_cash_program"),
                    *records_for("dividends_paid", 5),
                ],
            }
        }
    )
    mixed_records: list[dict] = []
    for fiscal_period, period_type, period_start, period_end in periods:
        buyback = _record(
            "repurchase_cash_program",
            25,
            unit="$m",
            currency="USD",
            fiscal_period=fiscal_period,
            period_type=period_type,
            period_start=period_start,
            period_end=period_end,
        )
        dividend = _record(
            "dividends_paid",
            5,
            unit="$m",
            currency="USD",
            fiscal_period=fiscal_period,
            period_type=period_type,
            period_start=period_start,
            period_end=period_end,
        )
        fcf = _record(
            "free_cash_flow",
            100,
            unit="$m",
            currency="USD",
            fiscal_period=fiscal_period,
            period_type=period_type,
            period_start=period_start,
            period_end=period_end,
        )
        total = derive_total_capital_return(buyback, dividend)
        mixed_records.extend(
            (
                buyback,
                dividend,
                fcf,
                total,
                derive_fcf_coverage(buyback, fcf, metric_id="buybacks_to_fcf"),
                derive_fcf_coverage(dividend, fcf, metric_id="dividends_to_fcf"),
                derive_fcf_coverage(total, fcf, metric_id="total_capital_return_to_fcf"),
            )
        )
    mixed = build_capital_return_workbook_projection(
        {
            "capital_returns": {
                "collection_version": "1.0.0",
                "records": mixed_records,
            }
        }
    )

    def product_value(projection, row_key: str) -> float | None:
        row = next(row for row in projection.product_rows if row["row_key"] == row_key)
        return row["latest_quarter"]

    assert product_value(buyback_only, "repurchase_cash_program") == 25
    assert product_value(buyback_only, "dividends_paid") is None
    assert product_value(dividend_only, "repurchase_cash_program") is None
    assert product_value(dividend_only, "dividends_paid") == 5
    assert product_value(mixed, "total_capital_return") == 30
    assert product_value(mixed, "buybacks_to_fcf") == 0.25
    assert product_value(mixed, "dividends_to_fcf") == 0.05
    assert product_value(mixed, "total_capital_return_to_fcf") == 0.30


def test_incomplete_ttm_and_duplicate_period_identity_fail_closed() -> None:
    quarter = _record("repurchase_cash_program", 10, unit="$m", currency="USD")
    annual = _record(
        "repurchase_cash_program",
        40,
        unit="$m",
        currency="USD",
        fiscal_period="2025-FY",
        period_type="annual",
        period_start="2025-02-02",
        period_end="2026-01-31",
    )
    with pytest.raises(CapitalReturnResolutionError, match="no 'ttm' period"):
        build_capital_return_workbook_projection(
            {
                "capital_returns": {
                    "collection_version": "1.0.0",
                    "records": [quarter, annual],
                }
            }
        )

    duplicate = deepcopy(quarter)
    duplicate["record_id"] = f"{quarter['record_id']}_conflict"
    with pytest.raises(CapitalReturnResolutionError, match="Duplicate Capital Return metric/period"):
        validate_capital_return_records([quarter, duplicate])


def test_valuation_capital_return_shell_is_bounded_and_hides_lineage() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation"
    _configure_valuation_capital_return_product(wb)

    assert ws["A152"].value == "Capital Return"
    assert [ws[f"{column}153"].value for column in "ABCDE"] == [
        "Metric",
        None,
        None,
        None,
        "State / context",
    ]
    assert all(not ws.row_dimensions[row].hidden for row in range(152, 169))
    assert all(ws.column_dimensions[column].hidden for column in (
        "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO"
    ))
    assert "A152:M152" in {str(value) for value in ws.merged_cells.ranges}
    assert "E154:M154" in {str(value) for value in ws.merged_cells.ranges}
    assert all("[Red]" not in ws[f"B{row}"].number_format for row in range(154, 169))
    assert all(ws[f"A{row}"].value is None for row in range(154, 169))
    assert ws["AD171"].value == "row_key"
    assert ws["A154"].font.name == "Aptos"
    assert ws["B154"].alignment.horizontal == "left"
    assert ws.row_dimensions[153].height == 42


def test_generic_runtime_has_no_ticker_branch() -> None:
    source = (
        Path(__file__).resolve().parents[1]
        / "pbi_xbrl"
        / "new_ticker_capital_return.py"
    ).read_text(encoding="utf-8")
    assert '"ANF"' not in source
    assert "'ANF'" not in source
