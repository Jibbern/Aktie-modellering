from __future__ import annotations

import datetime as dt
import time
from pathlib import Path
from typing import Iterable, Sequence

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName

from pbi_xbrl.excel_writer_context import (
    _history_q_latest_full_year_actuals_from_workbook,
    _history_q_latest_full_year_period_set,
)
from pbi_xbrl.workbook_validation_runner import validate_workbook


YELLOW_INPUT_FILL = "FFF2CC"


def _add_named_ranges(wb: Workbook) -> None:
    wb["Valuation"]["A1"] = 0.12
    for name in ["CompanyOperatingMargin_Latest", "OperatingMargin_Latest", "CompanyOperatingMargin_TTM"]:
        wb.defined_names.add(DefinedName(name=name, attr_text="'Valuation'!$A$1"))


def _required_sheets(wb: Workbook, ticker: str) -> None:
    if wb.active is not None:
        wb.active.title = "Valuation"
    for sheet_name in [
        f"{ticker}_Investment_Case",
        "Promise_Progress_UI",
        "Quarter_Notes_UI",
        "History_Q",
        "Operating_Drivers",
        "Needs_Review",
        "QA_Log",
        "QA_Checks",
        "Scenario_Bridge_Tax_Treatment",
        "Scenario_Driver_Assumptions",
        "Quarter_Narrative_Data",
        "BS_Segments",
    ]:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
    _add_named_ranges(wb)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb["Needs_Review"].append(["priority", "issue"])
    wb["Needs_Review"].append(["INFO", "dry-run fixture"])
    wb["QA_Log"].append(["check", "status"])
    wb["QA_Log"].append(["dry-run", "pass"])


def _add_history_q(wb: Workbook, rows: Sequence[tuple[dt.date, float, float, float, float, float]]) -> None:
    ws = wb["History_Q"]
    ws.append(["quarter", "fiscal_year", "fiscal_quarter", "revenue", "op_income", "eps", "fcf", "buybacks"])
    for qd, revenue_m, op_income_m, eps, fcf_m, buybacks_m in rows:
        fiscal_quarter = ((qd.month - 1) // 3) + 1
        ws.append([dt.datetime(qd.year, qd.month, qd.day), qd.year, fiscal_quarter, revenue_m * 1_000_000, op_income_m * 1_000_000, eps, fcf_m * 1_000_000, buybacks_m * 1_000_000])


def _add_manual_inputs(wb: Workbook, ticker: str, *, guidance: bool = False, segment: bool = False) -> None:
    ws = wb[f"{ticker}_Investment_Case"]
    ws.append(["Manual Market / Scenario Inputs"])
    ws.append(["Input", "Model default (latest full-year)", "Model default (TTM)", "Guidance (current year)", "Guidance (next quarter)", "Manual override", "Active value", "Notes"])
    rows = [
        ("Forward revenue", 1000.0, 1040.0, 1120.0 if guidance else "", "", "", '=IF(F3<>"",F3,IF(C3<>"",C3,IF(B3<>"",B3,IF(D3<>"",D3,E3))))', "Default actuals; guidance only when clean."),
        ("Forward EPS", 1.20, 1.30, 1.45 if guidance else "", "", "", '=IF(F4<>"",F4,IF(C4<>"",C4,IF(B4<>"",B4,IF(D4<>"",D4,E4))))', "EPS default."),
        ("Forward FCF", 80.0, 88.0, 95.0 if guidance else "", "", "", '=IF(F5<>"",F5,IF(C5<>"",C5,IF(B5<>"",B5,IF(D5<>"",D5,E5))))', "FCF default."),
        ("Operating margin", 0.12, 0.13, 0.14 if guidance else "", "", "", '=IF(F6<>"",F6,IF(C6<>"",C6,IF(B6<>"",B6,IF(D6<>"",D6,E6))))', "Operating income / revenue."),
        ("Scenario tax rate", 0.25, 0.25, "", "", "", '=IF(F7<>"",F7,IF(C7<>"",C7,IF(B7<>"",B7,IF(D7<>"",D7,IF(E7<>"",E7,0.25)))))', "25% default scenario tax rate if unavailable."),
    ]
    for row in rows:
        ws.append(list(row))
        ws.cell(ws.max_row, 6).fill = PatternFill("solid", fgColor=YELLOW_INPUT_FILL)
    if segment:
        ws.append([])
        ws.append(["Segment Scenario Inputs"])
        ws.append(["Segment / category", "Type", "Baseline revenue", "Revenue % change", "Revenue impact", "Operating margin", "EBITDA impact", "Feeds bridge?", "Notes"])
        ws.append(["Segment A", "Segment", 600.0, "", '=IF(D11="",0,C11*D11)', 0.20, "=E11*F11", "Yes", "Segment operating margin"])
        ws.append(["Segment B", "Segment", 400.0, "", '=IF(D12="",0,C12*D12)', 0.13, "=E12*F12", "Yes", "Company operating margin proxy"])
    ws.append([])
    ws.append(["Scenario Driver Bridge"])
    ws.append(["Bridge item", "Baseline included", "Active / guide", "Incremental effect", "EPS impact", "EBITDA impact", "FCF impact", "Read"])
    ws.append(["Selected segment revenue/margin impact" if segment else "Manual operating uplift", 0, "=SUMIF(H11:H12,\"Yes\",G11:G12)" if segment else 0, "=C15-B15", "=D15*(1-$G$7)/50", "=D15", 0, "Taxable operating uplift."])


def _add_promise_progress(wb: Workbook, *, guidance_heavy: bool = False) -> None:
    ws = wb["Promise_Progress_UI"]
    ws.append(["Promise Progress"])
    if not guidance_heavy:
        ws.append(["No clean guidance"])
        return
    ws.append(["2026 open guidance"])
    ws.append(["Metric", "Current guide", "Horizon", "Status", "Notes/source"])
    ws.append(["Revenue guidance", "$1.1bn-$1.2bn", "2026 year", "Open", "Initial 2026 annual guide."])
    ws.append([])
    ws.append(["2025-Q4 revisions"])
    ws.append(["Metric", "Previous guide", "New/current guide", "Change type", "Actual", "Progress / run-rate", "Status", "Horizon", "Stated in", "Source date", "Source / note"])
    ws.append(["Revenue guidance", "$1.0bn-$1.1bn", "$1.04bn actual", "Completed", "$1.04bn", "", "Hit", "2025 year", "2025-Q4", "2026-02-20", "Final annual result."])
    ws.append(["Revenue guidance", "", "$1.1bn-$1.2bn", "Initial", "", "", "Open", "2026 year", "2025-Q4", "2026-02-20", "Future annual guide remains open."])


def _save_fixture(tmp_path: Path, ticker: str, *, guidance: bool = False, segment: bool = False) -> Path:
    wb = Workbook()
    _required_sheets(wb, ticker)
    _add_history_q(
        wb,
        [
            (dt.date(2025, 3, 31), 240, 24, 0.25, 18, 0),
            (dt.date(2025, 6, 30), 250, 27, 0.28, 20, 0),
            (dt.date(2025, 9, 30), 260, 31, 0.32, 23, 0),
            (dt.date(2025, 12, 31), 290, 38, 0.45, 27, 5),
        ],
    )
    _add_manual_inputs(wb, ticker, guidance=guidance, segment=segment)
    _add_promise_progress(wb, guidance_heavy=guidance)
    wb["Quarter_Notes_UI"]["A1"] = "No fake narrative"
    wb["Operating_Drivers"]["A1"] = "Operating drivers"
    wb["Scenario_Bridge_Tax_Treatment"].append(["Ticker", "Bridge item", "Driver type", "Tax treatment"])
    wb["Scenario_Driver_Assumptions"].append(["Ticker", "Section", "Segment / category", "Feeds bridge?"])
    path = tmp_path / f"{ticker}_model.xlsx"
    wb.save(path)
    return path


def test_calendar_sparse_reporter_dry_run_uses_defaults_without_fake_guidance(tmp_path: Path) -> None:
    started = time.perf_counter()
    path = _save_fixture(tmp_path, "FAKECAL", guidance=False, segment=False)
    elapsed = time.perf_counter() - started

    result = validate_workbook(path, "FAKECAL")
    assert result.formula_error_count == 0
    assert result.needs_review_p1_count == 0
    assert elapsed < 5.0

    wb = Workbook()
    _required_sheets(wb, "FAKECAL2")
    _add_history_q(
        wb,
        [
            (dt.date(2025, 3, 31), 100, 10, 0.1, 5, 0),
            (dt.date(2025, 6, 30), 100, 10, 0.1, 5, 0),
            (dt.date(2025, 9, 30), 100, 10, 0.1, 5, 0),
            (dt.date(2025, 12, 31), 100, 10, 0.1, 5, 0),
        ],
    )
    period_set = _history_q_latest_full_year_period_set(wb, ticker="FAKECAL2")
    assert period_set["labels"] == ["2025-Q1", "2025-Q2", "2025-Q3", "2025-Q4"]


def test_non_calendar_fiscal_reporter_dry_run_uses_fiscal_profile_not_anf_hardcode() -> None:
    wb = Workbook()
    _required_sheets(wb, "FUTURE")
    _add_history_q(
        wb,
        [
            (dt.date(2024, 5, 4), 90, 9, 0.09, 4, 0),
            (dt.date(2024, 8, 3), 95, 10, 0.10, 5, 0),
            (dt.date(2024, 11, 2), 100, 11, 0.11, 5, 0),
            (dt.date(2025, 2, 1), 115, 15, 0.15, 8, 1),
            (dt.date(2025, 5, 3), 110, 13, 0.12, 7, 1),
            (dt.date(2025, 8, 2), 120, 15, 0.14, 8, 1),
            (dt.date(2025, 11, 1), 130, 18, 0.16, 9, 1),
            (dt.date(2026, 1, 31), 140, 21, 0.20, 10, 1),
        ],
    )
    ws = wb["History_Q"]
    for row_idx, fiscal_year, fiscal_quarter in [
        (2, 2024, 1),
        (3, 2024, 2),
        (4, 2024, 3),
        (5, 2024, 4),
        (6, 2025, 1),
        (7, 2025, 2),
        (8, 2025, 3),
        (9, 2025, 4),
    ]:
        ws.cell(row_idx, 2).value = fiscal_year
        ws.cell(row_idx, 3).value = fiscal_quarter

    period_set = _history_q_latest_full_year_period_set(
        wb,
        ticker="FUTURE_RETAIL",
        fiscal_profile={"year_end_month": 1, "year_end_day": 31, "year_label": "start"},
    )
    actuals = _history_q_latest_full_year_actuals_from_workbook(
        wb,
        ticker="FUTURE_RETAIL",
        fiscal_profile={"year_end_month": 1, "year_end_day": 31, "year_label": "start"},
    )

    assert period_set["fiscal_year"] == 2025
    assert period_set["quarter_dates"] == [
        dt.date(2025, 5, 3),
        dt.date(2025, 8, 2),
        dt.date(2025, 11, 1),
        dt.date(2026, 1, 31),
    ]
    assert actuals["revenue_m"] == pytest.approx(500.0)
    assert actuals["revenue_growth"] == pytest.approx((500.0 / 400.0) - 1.0)


def test_segment_reporter_dry_run_uses_segment_margin_and_labeled_proxy(tmp_path: Path) -> None:
    path = _save_fixture(tmp_path, "SEGCO", guidance=False, segment=True)
    result = validate_workbook(path, "SEGCO")
    assert result.formula_error_count == 0
    assert result.needs_review_p1_count == 0

    # The fixture models one source-backed segment margin and one explicit proxy; missing margins would not feed.
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False)
    try:
        ws = wb["SEGCO_Investment_Case"]
        assert ws["F11"].value == 0.20
        assert ws["I11"].value == "Segment operating margin"
        assert ws["I12"].value == "Company operating margin proxy"
        bridge_row = next(rr for rr in range(1, ws.max_row + 1) if ws.cell(rr, 1).value == "Selected segment revenue/margin impact")
        assert "SUMIF" in str(ws.cell(bridge_row, 3).value)
    finally:
        wb.close()


def test_guidance_heavy_reporter_keeps_future_annual_guidance_open(tmp_path: Path) -> None:
    path = _save_fixture(tmp_path, "GUIDECO", guidance=True, segment=False)
    result = validate_workbook(path, "GUIDECO")
    assert result.formula_error_count == 0
    assert result.needs_review_p1_count == 0

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    try:
        rows = list(wb["Promise_Progress_UI"].iter_rows(values_only=True))
        future_rows = [row for row in rows if row and "2026 year" in row]
        assert future_rows
        assert all("Completed" not in row and "Hit" not in row and "Missed" not in row for row in future_rows)
    finally:
        wb.close()
