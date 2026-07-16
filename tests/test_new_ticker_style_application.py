from __future__ import annotations

from copy import copy

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
import pytest

from pbi_xbrl.new_ticker_style_application import StyleApplicationError, apply_style_plan
from pbi_xbrl.new_ticker_style_planner import PlannedStyleAction, StylePlan


def _action(cell: str = "B2", color: str = "2F80ED") -> PlannedStyleAction:
    return PlannedStyleAction(
        sheet="Valuation",
        cell=cell,
        style_key=f"fixture|{cell}",
        policy_id="fixture_policy",
        period="2025-Q4",
        current_value=120.0,
        comparison_period="2024-Q4",
        comparison_value=100.0,
        signal_value=0.2,
        signal_band="strong_positive",
        overlay={"fill": {"fill_type": "solid", "fg_color": color}},
        lineage=("fixture:current", "fixture:prior"),
    )


def _plan(*actions: PlannedStyleAction) -> StylePlan:
    return StylePlan(
        ticker="TEST",
        module_profile_id="core_only",
        style_contract_digest="a" * 64,
        binding_plan_digest="b" * 64,
        actions=list(actions),
        decisions=[],
    )


def test_apply_style_plan_changes_only_the_declared_fill() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Valuation"
    cell = worksheet["B2"]
    cell.value = "=1+1"
    cell.font = Font(name="Aptos", size=11, bold=True, color="112233")
    cell.border = Border(bottom=Side(style="thin", color="445566"))
    cell.alignment = Alignment(horizontal="right", wrap_text=True)
    cell.protection = Protection(locked=True)
    cell.number_format = "0.0%"
    before = {
        "value": cell.value,
        "font": copy(cell.font),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "protection": copy(cell.protection),
        "number_format": cell.number_format,
    }

    result = apply_style_plan(workbook, _plan(_action()))

    assert result.applied_action_count == 1
    assert result.styled_cells == ("Valuation!B2",)
    assert cell.fill == PatternFill(fill_type="solid", fgColor="2F80ED")
    assert cell.value == before["value"]
    assert cell.font == before["font"]
    assert cell.border == before["border"]
    assert cell.alignment == before["alignment"]
    assert cell.protection == before["protection"]
    assert cell.number_format == before["number_format"]


def test_apply_style_plan_rejects_duplicate_or_fabricated_actions() -> None:
    workbook = Workbook()
    workbook.active.title = "Valuation"

    with pytest.raises(StyleApplicationError, match="more than once"):
        apply_style_plan(workbook, _plan(_action(), _action()))
    with pytest.raises(StyleApplicationError, match="missing sheet"):
        apply_style_plan(workbook, _plan(PlannedStyleAction(**{**_action().to_dict(), "sheet": "Missing", "lineage": ("fixture",)})))
    with pytest.raises(StyleApplicationError, match="reproduced PASS StylePlan"):
        apply_style_plan(workbook, {"status": "PASS"})  # type: ignore[arg-type]
