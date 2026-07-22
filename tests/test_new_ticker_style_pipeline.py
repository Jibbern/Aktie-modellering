from __future__ import annotations

from copy import copy
import ctypes
import gc
import json
from pathlib import Path
import shutil
import sys
from typing import Any
import zipfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import pytest

from pbi_xbrl.excel_formula_serialization import inventory_xlsx_formula_xml
from pbi_xbrl.new_ticker_style_planner import load_style_policy_contract, reproduce_style_plan
from pbi_xbrl.new_ticker_value_filler import fill_standard_template_from_package
from pbi_xbrl.standard_template_formula_contract import validate_workbook_protection_contract
from pbi_xbrl.standard_template_shell_identity import verify_post_fill_structural_identity
from pbi_xbrl.workbook_modules import load_workbook_module_manifest


ROOT = Path(__file__).resolve().parents[1]
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"

_EXCEL_ERROR_VALUES = {
    (0x800A0000 + code) - (1 << 32)
    for code in (2000, 2007, 2015, 2023, 2029, 2036, 2042, 2043)
}


def _json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _package_path() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = parent / "StockModelData" / "outputs" / "stress_tests" / "ANF_new_ticker_engine" / "ANF_normalized_data_package.json"
        if candidate.exists():
            return candidate
    pytest.skip("ANF normalized style fixture is unavailable.")


@pytest.fixture(scope="module")
def filled_anf_style_workbook(tmp_path_factory: pytest.TempPathFactory) -> dict[str, Any]:
    output = tmp_path_factory.mktemp("anf-style-fill") / "ANF_style_pipeline.xlsx"
    package_path = _package_path()
    result = fill_standard_template_from_package(package_path, output_path=output)
    package = _json(package_path)
    binding_payload = _json(BINDING_MAP)
    manifest = _json(MANIFEST)
    modules = load_workbook_module_manifest()
    styles = load_style_policy_contract(module_payload=modules, binding_payload=binding_payload)
    value_plan, style_plan = reproduce_style_plan(
        package,
        binding_payload=binding_payload,
        manifest=manifest,
        shell_path=SHELL,
        module_payload=modules,
        style_contract=styles,
    )
    return {
        "output": output,
        "result": result,
        "package": package,
        "binding": binding_payload,
        "manifest": manifest,
        "modules": modules,
        "styles": styles,
        "value_plan": value_plan,
        "style_plan": style_plan,
    }


def _non_fill_style(cell: Any) -> dict[str, Any]:
    return {
        "font": copy(cell.font),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "protection": copy(cell.protection),
        "number_format": cell.number_format,
    }


def _conditional_formatting_signature(worksheet: Any) -> list[tuple[str, tuple[tuple[str, tuple[str, ...]], ...]]]:
    return [
        (
            str(key.sqref),
            tuple(
                (str(rule.type or ""), tuple(str(formula) for formula in (rule.formula or [])))
                for rule in worksheet.conditional_formatting[key]
            ),
        )
        for key in worksheet.conditional_formatting
    ]


def _excel_range_errors(worksheet: Any, target: str | None = None) -> list[str]:
    used = worksheet.Range(target) if target else worksheet.UsedRange
    row_count = int(used.Rows.Count)
    column_count = int(used.Columns.Count)
    first_row = int(used.Row)
    first_column = int(used.Column)
    values = used.Value2
    if row_count == 1 and column_count == 1:
        matrix = ((values,),)
    else:
        matrix = values
    errors: list[str] = []
    for row_offset, row_values in enumerate(matrix):
        normalized_row = row_values if isinstance(row_values, tuple) else (row_values,)
        for column_offset, value in enumerate(normalized_row):
            if isinstance(value, int) and value in _EXCEL_ERROR_VALUES:
                coordinate = f"{get_column_letter(first_column + column_offset)}{first_row + row_offset}"
                formula = worksheet.Cells(first_row + row_offset, first_column + column_offset).Formula2
                errors.append(f"{worksheet.Name}!{coordinate}:{value}:{formula}")
    return errors


def _formula2_values(value: Any) -> tuple[str, ...]:
    if isinstance(value, tuple):
        return tuple(str(item) for row in value for item in (row if isinstance(row, tuple) else (row,)))
    return (str(value),)


def _wait_for_owned_process_exit(process_id: int) -> None:
    synchronize = 0x00100000
    process_terminate = 0x0001
    wait_timeout = 0x00000102
    kernel32 = ctypes.windll.kernel32
    handle = kernel32.OpenProcess(synchronize | process_terminate, False, process_id)
    if not handle:
        return
    try:
        result = kernel32.WaitForSingleObject(handle, 10_000)
        if result == wait_timeout:
            assert kernel32.TerminateProcess(handle, 1)
            result = kernel32.WaitForSingleObject(handle, 10_000)
        assert result == 0
    finally:
        kernel32.CloseHandle(handle)


def test_public_filler_applies_exact_reproduced_style_plan_after_values(
    filled_anf_style_workbook: dict[str, Any],
) -> None:
    artifacts = filled_anf_style_workbook
    result = artifacts["result"]
    assert result.written_cell_count == 22_760
    assert result.styled_cell_count == 738

    shell = load_workbook(SHELL, data_only=False, read_only=False)
    filled = load_workbook(artifacts["output"], data_only=False, read_only=False)
    try:
        assert filled["Valuation"]["B9"].fill.fgColor.rgb[-6:] == "2F80ED"
        assert filled["Valuation"]["B70"].fill.fill_type is None
        assert (
            filled["Valuation"]["AA9"].fill.patternType,
            filled["Valuation"]["AA9"].fill.fgColor.rgb,
        ) == (
            shell["Valuation"]["AA9"].fill.patternType,
            shell["Valuation"]["AA9"].fill.fgColor.rgb,
        )
        assert filled["Valuation"]["X51"].fill.fgColor.rgb[-6:] == "FFEB9C"
        assert (
            filled["Valuation"]["X52"].fill.patternType,
            filled["Valuation"]["X52"].fill.fgColor.rgb,
        ) == (
            shell["Valuation"]["X52"].fill.patternType,
            shell["Valuation"]["X52"].fill.fgColor.rgb,
        )
        assert _non_fill_style(filled["Valuation"]["B9"]) == _non_fill_style(shell["Valuation"]["B9"])
        assert _conditional_formatting_signature(filled["Valuation"]) == _conditional_formatting_signature(
            shell["Valuation"]
        )
    finally:
        filled.close()
        shell.close()


def test_strict_post_fill_accepts_only_the_reproduced_style_plan(
    filled_anf_style_workbook: dict[str, Any],
) -> None:
    artifacts = filled_anf_style_workbook
    report = verify_post_fill_structural_identity(
        artifacts["output"],
        approved_shell_path=SHELL,
        manifest=artifacts["manifest"],
        binding_payload=artifacts["binding"],
        approved_plan=artifacts["value_plan"],
        normalized_package=artifacts["package"],
        module_payload=artifacts["modules"],
        style_contract=artifacts["styles"],
        approved_style_plan=artifacts["style_plan"],
    )

    assert report["status"] == "PASS", report["issues"][:10]
    assert report["reproduced_style_action_count"] == 738


def test_strict_post_fill_rejects_fabricated_hidden_value_state_fill(
    filled_anf_style_workbook: dict[str, Any], tmp_path: Path
) -> None:
    artifacts = filled_anf_style_workbook
    drifted = tmp_path / "fabricated-hidden-value-state-fill.xlsx"
    shutil.copyfile(artifacts["output"], drifted)
    workbook = load_workbook(drifted, data_only=False, read_only=False)
    try:
        workbook["Hidden_Value_Audit"]["F2"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
        workbook.save(drifted)
    finally:
        workbook.close()

    report = verify_post_fill_structural_identity(
        drifted,
        approved_shell_path=SHELL,
        manifest=artifacts["manifest"],
        binding_payload=artifacts["binding"],
        approved_plan=artifacts["value_plan"],
        normalized_package=artifacts["package"],
        module_payload=artifacts["modules"],
        style_contract=artifacts["styles"],
        approved_style_plan=artifacts["style_plan"],
    )

    assert report["status"] == "FAIL"
    assert "post_fill_protected_cell_drift" in {row["rule_id"] for row in report["issues"]}


def test_strict_post_fill_rejects_one_unplanned_style_mutation(
    filled_anf_style_workbook: dict[str, Any], tmp_path: Path
) -> None:
    artifacts = filled_anf_style_workbook
    drifted = tmp_path / "unplanned-style.xlsx"
    shutil.copyfile(artifacts["output"], drifted)
    workbook = load_workbook(drifted, data_only=False, read_only=False)
    try:
        workbook["Valuation"]["B70"].fill = PatternFill(fill_type="solid", fgColor="2F80ED")
        workbook.save(drifted)
    finally:
        workbook.close()

    report = verify_post_fill_structural_identity(
        drifted,
        approved_shell_path=SHELL,
        manifest=artifacts["manifest"],
        binding_payload=artifacts["binding"],
        approved_plan=artifacts["value_plan"],
        normalized_package=artifacts["package"],
        module_payload=artifacts["modules"],
        style_contract=artifacts["styles"],
        approved_style_plan=artifacts["style_plan"],
    )

    assert report["status"] == "FAIL"
    assert "post_fill_protected_cell_drift" in {row["rule_id"] for row in report["issues"]}


@pytest.mark.skipif(sys.platform != "win32", reason="Desktop Excel automation is Windows-only")
def test_swedish_excel_native_recalculation_preserves_formula_and_protection_contracts(
    filled_anf_style_workbook: dict[str, Any], tmp_path: Path
) -> None:
    win32com = pytest.importorskip("win32com.client")
    pythoncom = pytest.importorskip("pythoncom")
    win32process = pytest.importorskip("win32process")
    artifacts = filled_anf_style_workbook
    path = tmp_path / "ANF_excel_native_formula_protection.xlsx"
    shutil.copyfile(artifacts["output"], path)
    sheet_names = tuple(
        str(row["sheet"]).replace("{ticker}", "ANF")
        for row in artifacts["manifest"]["sheets"]
    )

    excel = None
    book = None
    valuation = None
    process_id: int | None = None
    ui_language_id: int | None = None
    pythoncom.CoInitialize()
    try:
        try:
            excel = win32com.DispatchEx("Excel.Application")
        except Exception as exc:
            pytest.skip(f"Desktop Excel automation could not start in this session: {exc}")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        try:
            _thread_id, process_id = win32process.GetWindowThreadProcessId(int(excel.Hwnd))
        except Exception:
            process_id = None
        try:
            ui_language_id = int(excel.LanguageSettings.LanguageID(2))
        except Exception:
            ui_language_id = None
        try:
            book = excel.Workbooks.Open(
                str(path.resolve()),
                UpdateLinks=0,
                ReadOnly=False,
                IgnoreReadOnlyRecommended=True,
                CorruptLoad=0,
            )
        except Exception as exc:
            pytest.skip(f"Desktop Excel automation could not open the isolated workbook: {exc}")

        assert not bool(book.HasVBProject)
        assert all(bool(book.Worksheets(name).ProtectContents) for name in sheet_names)
        excel.CalculateFullRebuild()

        valuation = book.Worksheets("Valuation")
        future_formula_errors: list[str] = []
        for row in (10, 15, 21, 25, 34, 39, 46, 49, 50, 63, 64, 65, 66, 67, 88, 89, 109, 111, 271):
            formulas = _formula2_values(valuation.Range(f"B{row}:M{row}").Formula2)
            assert all("MAXIFS(" in formula or "MINIFS(" in formula for formula in formulas)
            future_formula_errors.extend(_excel_range_errors(valuation, f"B{row}:M{row}"))
        summary_formulas = _formula2_values(book.Worksheets("Valuation_Summary").Range("H2:K2").Formula2)
        assert all("LET(" in formula for formula in summary_formulas)
        future_formula_errors.extend(_excel_range_errors(book.Worksheets("Valuation_Summary"), "H2:K2"))
        assert future_formula_errors == [], "\n".join(future_formula_errors)

        fail_closed_errors = [
            error
            for sheet_name, target in (
                ("Valuation", "N244"),
                ("Valuation", "N261"),
                ("ANF_Investment_Case", "B68"),
                *(("Hidden_Value_Recompute", f"{column}{row}") for row in (16, 17, 18, 32, 34, 63, 90) for column in ("AB", "AC", "AD")),
            )
            for error in _excel_range_errors(book.Worksheets(sheet_name), target)
        ]
        assert fail_closed_errors == [], "\n".join(fail_closed_errors)
        assert all(
            book.Worksheets(sheet_name).Range(target).Value in (None, "")
            for sheet_name, target in (
                ("Valuation", "N244"),
                ("Valuation", "N261"),
                ("ANF_Investment_Case", "B68"),
            )
        )

        checked_sheets = (
            "Valuation",
            "ANF_Investment_Case",
            "Valuation_Summary",
            "Hidden_Value_Base",
            "Hidden_Value_Audit",
            "Hidden_Value_Recompute",
            "Hidden_Value_Flags",
        )
        formula_errors = [
            error
            for sheet_name in checked_sheets
            for error in _excel_range_errors(book.Worksheets(sheet_name))
        ]
        name_errors = [error for error in formula_errors if ":-2146826259:" in error]
        assert name_errors == [], "\n".join(name_errors)

        book.Save()
        book.Close(SaveChanges=False)
        book = excel.Workbooks.Open(
            str(path.resolve()),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            CorruptLoad=0,
        )
        excel.CalculateFullRebuild()
        assert all(bool(book.Worksheets(name).ProtectContents) for name in sheet_names)
        assert all(
            _excel_range_errors(book.Worksheets("Valuation"), f"B{row}:M{row}") == []
            for row in (10, 15, 21, 25, 34, 39, 46, 49, 50, 63, 64, 65, 66, 67, 88, 89, 109, 111, 271)
        )
        assert _excel_range_errors(book.Worksheets("Valuation_Summary"), "H2:K2") == []
        assert all(
            _excel_range_errors(book.Worksheets(sheet_name), target) == []
            for sheet_name, target in (
                ("Valuation", "N244"),
                ("Valuation", "N261"),
                ("ANF_Investment_Case", "B68"),
                *(("Hidden_Value_Recompute", f"{column}{row}") for row in (16, 17, 18, 32, 34, 63, 90) for column in ("AB", "AC", "AD")),
            )
        )
        assert all(
            book.Worksheets(sheet_name).Range(target).Value in (None, "")
            for sheet_name, target in (
                ("Valuation", "N244"),
                ("Valuation", "N261"),
                ("ANF_Investment_Case", "B68"),
            )
        )
        book.Save()
    finally:
        valuation = None
        if book is not None:
            try:
                book.Close(SaveChanges=False)
            except Exception:
                pass
            book = None
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
            excel = None
        gc.collect()
        pythoncom.CoUninitialize()
        if process_id is not None:
            _wait_for_owned_process_exit(process_id)

    inventory = inventory_xlsx_formula_xml(path)
    assert inventory["cell_formula_count"] == 2213
    assert inventory["function_counts"]["MAXIFS"] == 324
    assert inventory["function_counts"]["MINIFS"] == 324
    assert inventory["function_counts"]["LET"] == 4
    assert inventory["let_local_occurrences"] == 204
    assert inventory["unprefixed_future_functions"] == {}
    assert inventory["unsupported_functions"] == {}
    assert inventory["malformed_expressions"] == []

    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        enabled = artifacts["manifest"]["module_profile"]["enabled_formula_ids"]
        assert validate_workbook_protection_contract(workbook, enabled) == []
    finally:
        workbook.close()

    with zipfile.ZipFile(path, "r") as package:
        names = set(package.namelist())
        assert "xl/vbaProject.bin" not in names
        assert not any(name.startswith("xl/externalLinks/") for name in names)
        assert not any("recovery" in name.casefold() for name in names)

    report = verify_post_fill_structural_identity(
        path,
        approved_shell_path=SHELL,
        manifest=artifacts["manifest"],
        binding_payload=artifacts["binding"],
        approved_plan=artifacts["value_plan"],
        normalized_package=artifacts["package"],
        module_payload=artifacts["modules"],
        style_contract=artifacts["styles"],
        approved_style_plan=artifacts["style_plan"],
        excel_native_roundtrip=True,
    )
    assert report["status"] == "PASS", {
        "ui_language_id": ui_language_id,
        "issues": report["issues"][:20],
    }
    assert report["accepted_excel_native_normalizations"] == ["layout"]
