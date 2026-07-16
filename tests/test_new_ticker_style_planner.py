from __future__ import annotations

from copy import deepcopy
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import pytest

from pbi_xbrl.json_schema_validation import load_json_strict, validate_json_schema
from pbi_xbrl.new_ticker_binding_planner import BindingPlan
from pbi_xbrl.new_ticker_style_planner import (
    DEFAULT_MODULE_MANIFEST,
    DEFAULT_STYLE_POLICY,
    EconomicPoint,
    FORMULA_ECONOMIC_SPECS,
    STYLE_PLAN_SCHEMA,
    STYLE_POLICY_SCHEMA,
    StylePlanningError,
    _FormulaEvaluator,
    _point_from_mapping,
    _style_decision,
    classify_signal_band,
    load_style_policy_contract,
    plan_style_actions,
    reproduce_style_plan,
    style_policy_ids_for_profile,
    validate_active_style_contract,
    validate_style_policy_contract,
)
from pbi_xbrl.standard_template_formula_contract import FormulaTargetContract, formula_target_contracts
from pbi_xbrl.workbook_modules import (
    build_profile_binding_payload,
    build_profile_shell_manifest,
    resolve_module_profile,
)
from scripts.build_standard_template_style_policy_audit import build_audit, render_markdown


ROOT = Path(__file__).resolve().parents[1]
MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
LEGACY_COLORS = {
    "strong_negative": "A63A00",
    "negative": "D55E00",
    "neutral": "DDDDDD",
    "positive": "9BD3F5",
    "strong_positive": "2F80ED",
}


def _json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _data_root() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = parent / "StockModelData"
        if candidate.exists():
            return candidate
    pytest.skip("StockModelData legacy/style oracle is unavailable.")


def _threshold() -> dict[str, Any]:
    contract = load_style_policy_contract()
    return next(row for row in contract["threshold_sets"] if row["threshold_set_id"] == "legacy_five_band_change")


def _policy(*, basis: str, lag: int, polarity: str, period_type: str = "quarter") -> dict[str, Any]:
    return {
        "policy_id": "test_policy",
        "comparison_basis": basis,
        "comparison_lag": lag,
        "polarity": polarity,
        "period_type": period_type,
        "accepted_units": ["$m"],
    }


def _point(value: float, unit: str = "$m") -> EconomicPoint:
    return EconomicPoint(value, unit, (f"fixture:{value}",))


def _independent_band(signal: float) -> str:
    if signal <= -0.15:
        return "strong_negative"
    if signal <= -0.05:
        return "negative"
    if signal < 0.05:
        return "neutral"
    if signal < 0.15:
        return "positive"
    return "strong_positive"


def _shift_quarter(period: str, lag: int) -> str:
    year, quarter = period.split("-Q")
    ordinal = int(year) * 4 + int(quarter) - 1 - lag
    return f"{ordinal // 4}-Q{ordinal % 4 + 1}"


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (-0.15, "strong_negative"),
        (math.nextafter(-0.15, math.inf), "negative"),
        (-0.05, "negative"),
        (math.nextafter(-0.05, math.inf), "neutral"),
        (math.nextafter(0.05, -math.inf), "neutral"),
        (0.05, "positive"),
        (math.nextafter(0.15, -math.inf), "positive"),
        (0.15, "strong_positive"),
    ],
)
def test_legacy_threshold_boundaries_are_exact(value: float, expected: str) -> None:
    assert classify_signal_band(value, _threshold()) == expected


def test_style_contract_schema_palette_and_authoritative_ownership() -> None:
    contract = load_json_strict(DEFAULT_STYLE_POLICY)
    modules = load_json_strict(DEFAULT_MODULE_MANIFEST)
    bindings = load_json_strict(BINDING_MAP)

    assert validate_json_schema(contract, load_json_strict(STYLE_POLICY_SCHEMA)) == []
    assert validate_style_policy_contract(contract, module_payload=modules, binding_payload=bindings) == []
    assert {key: row["fill"]["fg_color"] for key, row in contract["palette_tokens"].items()} == LEGACY_COLORS
    assert len(contract["policies"]) == 51
    assert len(contract["style_disabled"]) == 19


def test_unknown_and_incompatible_period_axes_fail_closed_with_target_context() -> None:
    contract = load_style_policy_contract()
    modules = load_json_strict(DEFAULT_MODULE_MANIFEST)
    bindings = load_json_strict(BINDING_MAP)

    missing = deepcopy(contract)
    policy = next(row for row in missing["policies"] if row["policy_id"] == "valuation_core_formula_ttm")
    policy["period_axis_id"] = "missing_axis"
    missing_issues = validate_style_policy_contract(missing, module_payload=modules, binding_payload=bindings)

    swapped = deepcopy(contract)
    quarterly = next(row for row in swapped["policies"] if row["policy_id"] == "valuation_core_formula_ttm")
    quarterly["period_axis_id"] = "bs_annual_financial_periods"
    annual = next(row for row in swapped["policies"] if row["policy_id"] == "annual_formula_higher")
    annual["period_axis_id"] = "valuation_quarterly_periods"
    swapped_issues = validate_style_policy_contract(swapped, module_payload=modules, binding_payload=bindings)

    assert any(
        "valuation_core_formula_ttm" in issue
        and "missing_axis" in issue
        and "quarter" in issue
        and "revenue_ttm" in issue
        and "Valuation!B10:M10" in issue
        for issue in missing_issues
    )
    assert any("bs_annual_financial_periods" in issue and "fiscal_year" in issue for issue in swapped_issues)
    assert any("valuation_quarterly_periods" in issue and "quarter" in issue for issue in swapped_issues)


def test_duplicate_selector_and_unknown_style_ownership_fail_closed() -> None:
    contract = load_style_policy_contract()
    modules = load_json_strict(DEFAULT_MODULE_MANIFEST)
    bindings = load_json_strict(BINDING_MAP)
    duplicate = deepcopy(contract)
    duplicate["policies"][1]["target_selectors"].append(
        deepcopy(duplicate["policies"][0]["target_selectors"][0])
    )
    duplicate["policies"][1]["owned_style_ids"] = ["unknown_style_range"]

    issues = validate_style_policy_contract(duplicate, module_payload=modules, binding_payload=bindings)

    assert any("Duplicate style target selector" in issue for issue in issues)
    assert any("unknown style ownership" in issue for issue in issues)


def test_threshold_gaps_overlaps_and_duplicate_band_ids_fail_closed() -> None:
    contract = load_style_policy_contract()
    modules = load_json_strict(DEFAULT_MODULE_MANIFEST)
    bindings = load_json_strict(BINDING_MAP)
    malformed = deepcopy(contract)
    bands = malformed["threshold_sets"][0]["bands"]
    bands[1]["minimum"] = -0.14
    bands[2]["band_id"] = bands[1]["band_id"]
    bands[3]["minimum_inclusive"] = False
    bands[2]["maximum_inclusive"] = False

    issues = validate_style_policy_contract(malformed, module_payload=modules, binding_payload=bindings)

    assert any("Duplicate band_id" in issue for issue in issues)
    assert any("gap or overlap" in issue for issue in issues)
    assert any("exactly one adjacent band" in issue for issue in issues)


def test_higher_lower_neutral_and_disabled_polarity() -> None:
    threshold = _threshold()
    current = _point(120.0)
    lookup = lambda _period: _point(100.0)

    higher, _ = _style_decision(
        sheet="Valuation", cell="B9", style_key="higher", policy=_policy(
            basis="prior_quarter", lag=1, polarity="higher_is_better"
        ), period="2025-Q4", current=current, comparison_lookup=lookup,
        threshold=threshold, palettes={key: {"fill": {"fill_type": "solid", "fg_color": color}} for key, color in LEGACY_COLORS.items()},
        missing_reason="missing",
    )
    lower, _ = _style_decision(
        sheet="Valuation", cell="B72", style_key="lower", policy=_policy(
            basis="prior_quarter", lag=1, polarity="lower_is_better"
        ), period="2025-Q4", current=_point(80.0), comparison_lookup=lookup,
        threshold=threshold, palettes={key: {"fill": {"fill_type": "solid", "fg_color": color}} for key, color in LEGACY_COLORS.items()},
        missing_reason="missing",
    )
    neutral_change, _ = _style_decision(
        sheet="Valuation", cell="B9", style_key="neutral", policy=_policy(
            basis="prior_quarter", lag=1, polarity="neutral_change"
        ), period="2025-Q4", current=current, comparison_lookup=lookup,
        threshold=threshold, palettes={key: {"fill": {"fill_type": "solid", "fg_color": color}} for key, color in LEGACY_COLORS.items()},
        missing_reason="missing",
    )
    disabled, decision = _style_decision(
        sheet="Valuation", cell="B70", style_key="disabled", policy=_policy(
            basis="disabled", lag=0, polarity="disabled"
        ), period="2025-Q4", current=current, comparison_lookup=lookup,
        threshold=threshold, palettes={}, missing_reason="missing",
    )

    assert higher is not None and higher.signal_band == "strong_positive"
    assert lower is not None and lower.signal_band == "strong_positive"
    assert neutral_change is not None and neutral_change.signal_band == "strong_positive"
    assert disabled is None and decision.reason == "policy_disabled"


def test_blank_missing_review_and_unit_mismatch_never_style() -> None:
    threshold = _threshold()
    palettes = {key: {"fill": {"fill_type": "solid", "fg_color": color}} for key, color in LEGACY_COLORS.items()}
    policy = _policy(basis="prior_quarter", lag=1, polarity="higher_is_better")

    blank, blank_decision = _style_decision(
        sheet="Valuation", cell="B9", style_key="blank", policy=policy, period="2025-Q4",
        current=None, comparison_lookup=lambda _period: _point(100), threshold=threshold,
        palettes=palettes, missing_reason="current_value_missing_or_untrusted",
    )
    mismatch, mismatch_decision = _style_decision(
        sheet="Valuation", cell="B9", style_key="unit", policy=policy, period="2025-Q4",
        current=_point(120), comparison_lookup=lambda _period: _point(100, "%"), threshold=threshold,
        palettes=palettes, missing_reason="missing",
    )

    assert blank is None and blank_decision.reason == "current_value_missing_or_untrusted"
    assert mismatch is None and mismatch_decision.reason == "comparator_unit_mismatch"
    assert _point_from_mapping({"value": 10, "unit": "$m", "source_ref": "fixture", "status": "missing_source"}) is None
    assert _point_from_mapping({"value": 10, "unit": "$m", "source_ref": "fixture", "status": "manual_review_required"}) is None


def test_ttm_requires_four_consecutive_source_backed_compatible_quarters() -> None:
    history = {
        "quarter": {
            "revenue": {
                "2025-Q1": _point(100),
                "2025-Q2": _point(110),
                "2025-Q3": _point(120),
                "2025-Q4": _point(130),
            }
        },
        "fiscal_year": {},
    }
    evaluator = _FormulaEvaluator(history, set())
    result, reason = evaluator.evaluate("revenue_ttm", "2025-Q4")
    assert result == EconomicPoint(460.0, "$m", ("fixture:100", "fixture:110", "fixture:120", "fixture:130"))
    assert reason == "calculated"

    missing = deepcopy(history)
    del missing["quarter"]["revenue"]["2025-Q2"]
    assert _FormulaEvaluator(missing, set()).evaluate("revenue_ttm", "2025-Q4")[1] == "formula_ttm_input_missing"

    mixed = deepcopy(history)
    mixed["quarter"]["revenue"]["2025-Q2"] = _point(110, "%")
    assert _FormulaEvaluator(mixed, set()).evaluate("revenue_ttm", "2025-Q4")[1] == "formula_ttm_unit_or_component_invalid"


def test_period_contracts_lock_qoq_yoy_ttm_and_annual_lags() -> None:
    contract = load_style_policy_contract()
    policies = {row["policy_id"]: row for row in contract["policies"]}

    assert (policies["valuation_buybacks_raw_qoq"]["period_type"], policies["valuation_buybacks_raw_qoq"]["comparison_lag"]) == ("quarter", 1)
    assert (policies["valuation_core_raw_yoy_higher"]["period_type"], policies["valuation_core_raw_yoy_higher"]["comparison_lag"]) == ("quarter", 4)
    assert (policies["valuation_core_formula_ttm"]["comparison_basis"], policies["valuation_core_formula_ttm"]["comparison_lag"]) == ("prior_ttm", 4)
    assert (policies["segment_annual_revenue"]["period_type"], policies["segment_annual_revenue"]["comparison_lag"]) == ("fiscal_year", 1)
    assert policies["segment_quarterly_revenue"]["comparison_lag"] == 4


def test_annual_signal_compares_immediately_preceding_fiscal_year() -> None:
    threshold = _threshold()
    palettes = {key: {"fill": {"fill_type": "solid", "fg_color": color}} for key, color in LEGACY_COLORS.items()}
    requested_periods: list[str] = []

    action, _decision_row = _style_decision(
        sheet="BS_Segments",
        cell="I72",
        style_key="annual-segment",
        policy=_policy(
            basis="prior_fiscal_year", lag=1, polarity="higher_is_better", period_type="fiscal_year"
        ),
        period="2025-FY",
        current=_point(110),
        comparison_lookup=lambda period: requested_periods.append(period) or _point(100),
        threshold=threshold,
        palettes=palettes,
        missing_reason="missing",
    )

    assert requested_periods == ["2024-FY"]
    assert action is not None and action.signal_band == "positive"


def test_formula_yoy_requires_matching_prior_year_metric_and_unit() -> None:
    history = {
        "quarter": {"revenue": {"2024-Q4": _point(100), "2025-Q4": _point(120)}},
        "fiscal_year": {},
    }
    result, reason = _FormulaEvaluator(history, set()).evaluate("revenue_yoy", "2025-Q4")
    assert reason == "calculated"
    assert result is not None and result.unit == "%" and result.value == pytest.approx(0.2)

    history["quarter"]["revenue"]["2024-Q4"] = _point(100, "%")
    assert _FormulaEvaluator(history, set()).evaluate("revenue_yoy", "2025-Q4")[1] == "formula_comparison_unit_mismatch"


def test_formula_economics_preserve_positive_capex_outflow_and_fcf_subtraction() -> None:
    history = {
        "quarter": {
            "operating_cash_flow": {"2025-Q4": _point(100)},
            "capital_expenditures": {"2025-Q4": _point(25)},
            "revenue": {"2025-Q4": _point(500)},
        },
        "fiscal_year": {},
    }
    evaluator = _FormulaEvaluator(history, set())

    fcf, fcf_reason = evaluator.evaluate("free_cash_flow", "2025-Q4")
    margin, margin_reason = evaluator.evaluate("free_cash_flow_margin", "2025-Q4")

    assert fcf_reason == "calculated" and fcf == EconomicPoint(75.0, "$m", ("fixture:100", "fixture:25"))
    assert margin_reason == "calculated" and margin is not None and margin.value == pytest.approx(0.15)


def test_profiles_activate_only_style_policies_owned_by_enabled_modules() -> None:
    contract = load_style_policy_contract()
    modules = load_json_strict(DEFAULT_MODULE_MANIFEST)
    core = style_policy_ids_for_profile(contract, modules, "core_only")
    anf = style_policy_ids_for_profile(contract, modules, "anf")

    assert "valuation_core_raw_yoy_higher" in core
    assert "valuation_debt_formula_lower" not in core
    assert "valuation_adjusted_raw_yoy" not in core
    assert "bs_debt_raw_lower_yoy" not in core
    assert "valuation_debt_formula_lower" in anf
    assert "valuation_adjusted_raw_yoy" in anf
    assert all(ticker not in " ".join(core).lower() for ticker in ("anf", "pbi", "gpre", "gtx"))


@pytest.fixture(scope="module")
def anf_style_artifacts() -> tuple[Any, Any]:
    root = _data_root()
    package_path = root / "outputs" / "stress_tests" / "ANF_new_ticker_engine" / "ANF_normalized_data_package.json"
    if not package_path.exists():
        pytest.skip("ANF normalized style fixture is unavailable.")
    return reproduce_style_plan(
        _json(package_path),
        binding_payload=_json(BINDING_MAP),
        manifest=_json(MANIFEST),
        shell_path=SHELL,
    )


def test_anf_style_plan_is_deterministic_and_value_plan_is_unchanged(anf_style_artifacts: tuple[Any, Any]) -> None:
    value_plan, style_plan = anf_style_artifacts

    assert value_plan.status == "PASS"
    assert len(value_plan.planned_writes) == 20_518
    assert len(style_plan.actions) == 824
    assert len(style_plan.decisions) == 1_372
    assert validate_json_schema(style_plan.to_dict(), load_json_strict(STYLE_PLAN_SCHEMA)) == []
    assert {action.sheet for action in style_plan.actions} == {"Valuation", "BS_Segments"}
    assert not any(action.cell in {"B70", "C70", "D70", "E70", "F70", "G70", "H70", "I70", "J70", "K70", "L70", "M70"} and action.sheet == "Valuation" for action in style_plan.actions)


def test_active_formula_style_inventory_is_exact_and_mutation_complete(
    anf_style_artifacts: tuple[Any, Any],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    value_plan, _style_plan = anf_style_artifacts
    contract = load_style_policy_contract()
    modules = load_json_strict(DEFAULT_MODULE_MANIFEST)
    bindings = load_json_strict(BINDING_MAP)
    manifest = load_json_strict(MANIFEST)

    assert validate_active_style_contract(
        contract,
        binding_plan=value_plan,
        binding_payload=bindings,
        manifest=manifest,
        module_payload=modules,
    ) == []

    removed = deepcopy(contract)
    share_policy = next(
        row for row in removed["policies"] if row["policy_id"] == "valuation_share_delta_direct_lower"
    )
    share_policy["target_selectors"] = [
        row for row in share_policy["target_selectors"] if row["target_id"] != "diluted_shares_qoq"
    ]
    removed_issues = validate_active_style_contract(
        removed,
        binding_plan=value_plan,
        binding_payload=bindings,
        manifest=manifest,
        module_payload=modules,
    )
    assert any(
        "diluted_shares_qoq" in issue and "no style selector or style_disabled disposition" in issue
        for issue in removed_issues
    )

    duplicate = deepcopy(contract)
    duplicate_policy = deepcopy(
        next(row for row in duplicate["policies"] if row["policy_id"] == "valuation_cash_flow_delta_direct")
    )
    duplicate_policy["policy_id"] = "valuation_cash_flow_delta_duplicate"
    duplicate["policies"].append(duplicate_policy)
    duplicate_issues = validate_active_style_contract(
        duplicate,
        binding_plan=value_plan,
        binding_payload=bindings,
        manifest=manifest,
        module_payload=modules,
    )
    assert any("free_cash_flow_yoy_delta" in issue and "duplicate style coverage" in issue for issue in duplicate_issues)

    original_targets = formula_target_contracts()
    monkeypatch.setattr(
        "pbi_xbrl.new_ticker_style_planner.formula_target_contracts",
        lambda: original_targets + (FormulaTargetContract("new_active_formula", "Valuation", ("B51:M51",)),),
    )
    mutated_modules = deepcopy(modules)
    balance_module = next(row for row in mutated_modules["modules"] if row["module_id"] == "balance_cash_flow")
    balance_module["formula_ids"].append("new_active_formula")
    mutated_manifest = deepcopy(manifest)
    mutated_manifest["module_profile"]["enabled_formula_ids"].append("new_active_formula")
    new_formula_issues = validate_active_style_contract(
        contract,
        binding_plan=value_plan,
        binding_payload=bindings,
        manifest=mutated_manifest,
        module_payload=mutated_modules,
    )
    assert any(
        "new_active_formula" in issue and "no style selector or style_disabled disposition" in issue
        for issue in new_formula_issues
    )


def test_missing_reproduced_axis_blocks_style_planning_before_decisions(
    anf_style_artifacts: tuple[Any, Any],
) -> None:
    value_plan, _style_plan = anf_style_artifacts
    mutated_plan = deepcopy(value_plan)
    mutated_plan.period_axes.pop("valuation_quarterly_periods")

    with pytest.raises(StylePlanningError, match="axis is absent from the independently reproduced plan"):
        plan_style_actions(
            {},
            binding_plan=mutated_plan,
            binding_payload=load_json_strict(BINDING_MAP),
            manifest=load_json_strict(MANIFEST),
            module_payload=load_json_strict(DEFAULT_MODULE_MANIFEST),
            style_contract=load_style_policy_contract(),
        )


def test_representative_valuation_styles_match_independent_legacy_oracle(anf_style_artifacts: tuple[Any, Any]) -> None:
    _value_plan, style_plan = anf_style_artifacts
    action_by_cell = {f"{row.sheet}!{row.cell}": row for row in style_plan.actions}
    expected = {
        "Valuation!B9": "2F80ED",
        "Valuation!B10": "DDDDDD",
        "Valuation!B11": "2F80ED",
        "Valuation!B14": "2F80ED",
        "Valuation!B18": "2F80ED",
        "Valuation!B20": "2F80ED",
        "Valuation!B21": "2F80ED",
        "Valuation!B43": "A63A00",
        "Valuation!B44": "A63A00",
        "Valuation!B47": "A63A00",
        "Valuation!B57": "A63A00",
        "Valuation!B72": "DDDDDD",
        "Valuation!B73": "2F80ED",
        "Valuation!B107": "2F80ED",
        "Valuation!B109": "2F80ED",
    }

    assert {cell: action_by_cell[cell].overlay["fill"]["fg_color"] for cell in expected} == expected
    legacy_path = _data_root() / "outputs" / "Excel stock models" / "ANF_model.xlsx"
    wb = load_workbook(legacy_path, read_only=False, data_only=False)
    try:
        assert {
            cell: wb["Valuation"][cell.split("!")[1]].fill.fgColor.rgb[-6:]
            for cell in expected
        } == expected
        assert wb["Valuation"]["B70"].fill.fill_type is None
    finally:
        wb.close()


def test_five_completed_formula_rows_follow_independent_legacy_and_business_rules(
    anf_style_artifacts: tuple[Any, Any],
) -> None:
    value_plan, style_plan = anf_style_artifacts
    actions = {f"{row.sheet}!{row.cell}": row for row in style_plan.actions}
    targeted = {
        key: row
        for key, row in actions.items()
        if row.sheet == "Valuation" and int(row.cell[1:]) in {48, 50, 90, 104, 105}
    }
    assert len(targeted) == 60

    legacy_path = _data_root() / "outputs" / "Excel stock models" / "ANF_model.xlsx"
    legacy = load_workbook(legacy_path, read_only=False, data_only=True)
    try:
        for row_number, lower_is_better in ((48, False), (50, False), (104, True), (105, True)):
            for column in "BCDEFGHIJKLM":
                value = float(legacy["Valuation"][f"{column}{row_number}"].value)
                expected_band = _independent_band(-value if lower_is_better else value)
                action = targeted[f"Valuation!{column}{row_number}"]
                assert action.comparison_period is None
                assert action.signal_band == expected_band
                assert action.overlay["fill"]["fg_color"] == LEGACY_COLORS[expected_band]
    finally:
        legacy.close()

    package_path = _data_root() / "outputs" / "stress_tests" / "ANF_new_ticker_engine" / "ANF_normalized_data_package.json"
    package = _json(package_path)
    history: dict[str, dict[str, float]] = {}
    for row in package["calculation_history"]["quarterly_items"]:
        if (
            row.get("metric") in {"operating_cash_flow", "capital_expenditures", "base_ebitda"}
            and row.get("status") in {"populated", "source_backed"}
            and row.get("source_ref")
            and row.get("unit") == "$m"
            and isinstance(row.get("value"), (int, float))
        ):
            history.setdefault(str(row["metric"]), {})[str(row["period"])] = float(row["value"])

    def conversion(period: str) -> float:
        periods = [_shift_quarter(period, lag) for lag in range(4)]
        cfo = sum(history["operating_cash_flow"][item] for item in periods)
        capex = sum(history["capital_expenditures"][item] for item in periods)
        base_ebitda = sum(history["base_ebitda"][item] for item in periods)
        return (cfo - capex) / base_ebitda

    axis = value_plan.period_axes["valuation_quarterly_periods"]
    for period, column in axis["period_to_column"].items():
        prior_period = _shift_quarter(period, 4)
        current = conversion(period)
        prior = conversion(prior_period)
        expected_band = _independent_band((current - prior) / abs(prior))
        action = targeted[f"Valuation!{column}90"]
        assert action.current_value == pytest.approx(current)
        assert action.comparison_period == prior_period
        assert action.comparison_value == pytest.approx(prior)
        assert action.signal_band == expected_band
        assert action.overlay["fill"]["fg_color"] == LEGACY_COLORS[expected_band]

    # The legacy fixture used adjusted EBITDA when available; the accepted
    # generic formula contract uses base EBITDA. That changes only the latest
    # signal from positive to neutral and preserves the accepted formula meaning.
    legacy_formula_view = load_workbook(legacy_path, read_only=False, data_only=False)
    try:
        assert legacy_formula_view["Valuation"]["M90"].fill.fgColor.rgb[-6:] == LEGACY_COLORS["positive"]
        assert targeted["Valuation!M90"].overlay["fill"]["fg_color"] == LEGACY_COLORS["neutral"]
    finally:
        legacy_formula_view.close()


def test_representative_bs_and_segment_styles_match_legacy_by_business_period(
    anf_style_artifacts: tuple[Any, Any],
) -> None:
    _value_plan, style_plan = anf_style_artifacts
    action_by_cell = {f"{row.sheet}!{row.cell}": row for row in style_plan.actions}
    # The union shell exposes more history than the legacy fixture, so compare
    # equivalent metric/period cells rather than identical column addresses.
    legacy_cells = {
        "BS_Segments!F9": "B9",   # cash, 2024-Q2
        "BS_Segments!I9": "E9",   # cash, 2025-Q1
        "BS_Segments!M9": "I9",   # cash, 2026-Q1
        "BS_Segments!G12": "C12", # cash QoQ delta, 2024-Q3
        "BS_Segments!L12": "H12", # cash QoQ delta, 2025-Q4
        "BS_Segments!F49": "B49", # diluted shares, 2024-Q2
        "BS_Segments!M49": "I49", # diluted shares, 2026-Q1
        "BS_Segments!F52": "B52", # revenue YoY, 2024-Q2
        "BS_Segments!M52": "I52", # revenue YoY, 2026-Q1
        "BS_Segments!F61": "B61", # Americas revenue, 2024-Q2
        "BS_Segments!G62": "C62", # EMEA revenue, 2024-Q3
        "BS_Segments!I63": "E63", # APAC revenue, 2025-Q1
        "BS_Segments!L65": "H65", # Total Company revenue, 2025-Q4
    }
    legacy_path = _data_root() / "outputs" / "Excel stock models" / "ANF_model.xlsx"
    wb = load_workbook(legacy_path, read_only=False, data_only=False)
    try:
        legacy = wb["BS_Segments"]
        assert {
            cell: action_by_cell[cell].overlay["fill"]["fg_color"]
            for cell in legacy_cells
        } == {
            cell: legacy[legacy_cell].fill.fgColor.rgb[-6:]
            for cell, legacy_cell in legacy_cells.items()
        }

        # Legacy annual segment coloring used a quarterly four-column lag.
        # The generic contract intentionally compares the immediately prior FY.
        annual = action_by_cell["BS_Segments!H72"]
        assert annual.period == "2024-FY"
        assert annual.comparison_period == "2023-FY"
        assert annual.current_value == pytest.approx(4027.5)
        assert annual.comparison_value == pytest.approx(3455.7)
        assert annual.signal_band == "strong_positive"
        assert annual.overlay["fill"]["fg_color"] == "2F80ED"
        assert legacy["C72"].fill.fgColor.rgb[-6:] == "DDDDDD"
    finally:
        wb.close()


def test_all_formula_selectors_with_active_signals_have_economic_specs() -> None:
    contract = load_style_policy_contract()
    required = {
        selector["target_id"]
        for policy in contract["policies"]
        if policy["comparison_basis"] != "disabled"
        for selector in policy["target_selectors"]
        if selector["selector_type"] == "formula"
    }

    assert required <= set(FORMULA_ECONOMIC_SPECS)


def test_checked_in_style_audit_is_an_exact_generated_projection() -> None:
    audit = build_audit()

    assert audit == _json(ROOT / "docs" / "standard_template_style_policy_audit.json")
    assert render_markdown(audit) == (ROOT / "docs" / "standard_template_style_policy_audit.md").read_text(
        encoding="utf-8"
    )


def test_all_declared_ticker_profiles_resolve_style_policies_deterministically() -> None:
    contract = load_style_policy_contract()
    modules = load_json_strict(DEFAULT_MODULE_MANIFEST)
    projections = {
        profile: style_policy_ids_for_profile(contract, modules, profile)
        for profile in ("anf", "pbi", "gpre", "core_only")
    }

    assert projections["anf"] == projections["pbi"] == projections["gpre"]
    assert set(projections["core_only"]) < set(projections["anf"])
    assert projections == {
        profile: style_policy_ids_for_profile(contract, modules, profile)
        for profile in projections
    }


def test_runtime_style_planning_uses_only_each_resolved_profile_modules() -> None:
    contract = load_style_policy_contract()
    modules = load_json_strict(DEFAULT_MODULE_MANIFEST)
    base_bindings = load_json_strict(BINDING_MAP)
    base_manifest = load_json_strict(MANIFEST)
    points = [
        {
            "period": f"{year}-Q{quarter}",
            "metric": "revenue",
            "value": 100.0 + index * 10.0,
            "unit": "$m",
            "status": "source_backed",
            "source_ref": f"fixture:{year}-Q{quarter}",
        }
        for index, (year, quarter) in enumerate(
            [(2024, 1), (2024, 2), (2024, 3), (2024, 4), (2025, 1), (2025, 2), (2025, 3), (2025, 4)]
        )
    ]
    package = {"calculation_history": {"quarterly_items": points}, "annual_financials": {"rows": []}, "segments": {"items": []}}
    policy_owner = {row["policy_id"]: row["owner_module_id"] for row in contract["policies"]}

    for profile_id in ("anf", "pbi", "gpre", "core_only"):
        resolved = resolve_module_profile(modules, profile_id)
        bindings = build_profile_binding_payload(base_bindings, modules, resolved)
        manifest = build_profile_shell_manifest(base_manifest, modules, resolved)
        plan = BindingPlan(
            ticker="TEST",
            period_axes={
                "valuation_quarterly_periods": {"period_to_column": {"2025-Q4": "B"}},
                "bs_quarterly_periods": {"period_to_column": {"2025-Q4": "B"}},
                "bs_annual_periods": {"period_to_column": {"2025-FY": "B"}},
                "bs_annual_financial_periods": {"period_to_column": {"2025-FY": "B"}},
            },
        )
        style_plan = plan_style_actions(
            package,
            binding_plan=plan,
            binding_payload=bindings,
            manifest=manifest,
            module_payload=modules,
            style_contract=contract,
        )

        assert style_plan.actions
        assert {policy_owner[action.policy_id] for action in style_plan.actions} <= set(resolved.enabled_modules)
        if profile_id == "core_only":
            assert not {
                "debt_liquidity",
                "non_gaap_adjustments",
                "segments_dimensions",
            } & {policy_owner[action.policy_id] for action in style_plan.actions}
