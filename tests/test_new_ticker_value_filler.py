from __future__ import annotations

from copy import deepcopy
from io import BytesIO
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import pytest

from pbi_xbrl.new_ticker_binding_planner import (
    BindingPlan,
    reproduce_binding_plan_snapshot,
    plan_standard_template_writes,
)
from pbi_xbrl.new_ticker_value_filler import (
    BindingContractError,
    NormalizedDataValidationError,
    _execute_binding_plan,
    _resolve_ticker_sheet,
    fill_standard_template_from_package,
)
from pbi_xbrl.standard_template_shell_identity import verify_shell_identity


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "standard_stock_model_template.xlsx"
MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
PACKAGE = ROOT / "tests" / "fixtures" / "normalized_packages" / "TEST_minimal_valid.json"


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _approved_plan(package: dict[str, Any] | None = None) -> BindingPlan:
    manifest = _load_json(MANIFEST)
    bindings = _load_json(BINDING_MAP)
    identity = verify_shell_identity(TEMPLATE, manifest=manifest, binding_payload=bindings)
    assert identity.status == "PASS", identity.issues
    return plan_standard_template_writes(
        package or _load_json(PACKAGE),
        binding_payload=bindings,
        manifest=manifest,
        shell_identity_report=identity,
    )


def _execute_in_memory(package: dict[str, Any] | None = None):
    normalized = package or _load_json(PACKAGE)
    manifest = _load_json(MANIFEST)
    bindings = _load_json(BINDING_MAP)
    identity = verify_shell_identity(TEMPLATE, manifest=manifest, binding_payload=bindings)
    plan, _snapshot = reproduce_binding_plan_snapshot(
        normalized,
        binding_payload=bindings,
        manifest=manifest,
        shell_path=TEMPLATE,
        shell_identity_report=identity,
    )
    workbook = load_workbook(BytesIO(TEMPLATE.read_bytes()), data_only=False, read_only=False)
    _resolve_ticker_sheet(workbook, "TEST")
    written = _execute_binding_plan(workbook, plan)
    return workbook, plan, written


def _formula_map(workbook) -> dict[tuple[str, str], str]:
    formulas: dict[tuple[str, str], str] = {}
    for ws in workbook.worksheets:
        sheet_name = ws.title.replace("{ticker}", "TEST")
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell) and isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas[(sheet_name, cell.coordinate)] = cell.value
    return formulas


def _layout_signature(workbook) -> dict[str, Any]:
    return {
        "sheet_order": [name.replace("{ticker}", "TEST") for name in workbook.sheetnames],
        "sheets": {
            ws.title.replace("{ticker}", "TEST"): {
                "state": ws.sheet_state,
                "freeze_panes": str(ws.freeze_panes or ""),
                "merges": sorted(str(item) for item in ws.merged_cells.ranges),
                "row_heights": {str(idx): dim.height for idx, dim in ws.row_dimensions.items() if dim.height is not None},
                "column_widths": {key: dim.width for key, dim in ws.column_dimensions.items() if dim.width is not None},
                "hidden_columns": {key: bool(dim.hidden) for key, dim in ws.column_dimensions.items() if dim.hidden},
            }
            for ws in workbook.worksheets
        },
    }


def test_exact_cell_plan_executes_in_memory_without_obsolete_targets() -> None:
    workbook, plan, written = _execute_in_memory()
    try:
        assert written == len(plan.planned_writes)
        assert workbook["SUMMARY"]["B26"].value == "2026-Q1"
        assert workbook["SUMMARY"]["B28"].value == 120.5
        assert workbook["SUMMARY"]["B30"].value == 13.1
        assert workbook["Valuation"]["B6"].value == "2025-Q4"
        assert workbook["Valuation"]["B9"].value == 118.0
        assert workbook["BS_Segments"]["B7"].value == "2025-Q4"
        assert workbook["BS_Segments"]["C66"].value == 92.0

        assert workbook["Operating_Drivers"]["A6"].value == "Demand"
        assert workbook["Operating_Drivers"]["B6"].value.startswith("Renewal rates remain stable")
        assert workbook["Operating_Drivers"]["H6"].value.startswith("Retention and implementation conversion")
        assert "topic:" not in str(workbook["Operating_Drivers"]["B6"].value)

        assert workbook["Quarter_Notes_UI"]["A9"].value == "Theme"
        assert workbook["Quarter_Notes_UI"]["A10"].value == "Demand"
        assert workbook["Quarter_Notes_UI"]["C10"].value.startswith("Recurring software demand")
        assert workbook["Quarter_Notes_UI"]["H10"].value.startswith("Growth durability")
        assert workbook["Quarter_Notes_UI"]["M10"].value == "synthetic_fixture:quarter_notes"

        assert workbook["Promise_Progress_UI"]["A61"].value == "Revenue"
        assert workbook["Promise_Progress_UI"]["C61"].value.startswith("Revenue growth")
        assert workbook["Promise_Progress_UI"]["G61"].value == "open"
        assert workbook["Promise_Progress_UI"]["I61"].value == "2026-Q1"
        assert workbook["Promise_Progress_UI"]["J61"].value == "2026-07-07"
        assert workbook["TEST_Investment_Case"]["B5"].value.startswith("The test case depends")
    finally:
        workbook.close()


def test_row_schema_values_stay_in_distinct_exact_cells() -> None:
    package = deepcopy(_load_json(PACKAGE))
    source_ref = "synthetic_fixture:row_schema"
    package["normalized_guidance"]["items"][0].update(
        {
            "initial_guide": {"value": "Initial guide", "status": "populated", "source_ref": source_ref, "core": False},
            "q1_update": {"value": "Q1 update", "status": "populated", "source_ref": source_ref, "core": False},
            "actual": {"value": "Actual result", "status": "populated", "source_ref": source_ref, "core": False},
            "progress_status": "Open",
            "notes_source": "Source-backed note",
        }
    )
    package["quarter_notes"]["items"][0]["source"] = source_ref
    package["operating_drivers"]["items"][0]["source"] = source_ref

    workbook, _plan, _written = _execute_in_memory(package)
    try:
        assert workbook["Operating_Drivers"]["A6"].value == "Demand"
        assert workbook["Operating_Drivers"]["B6"].value.startswith("Renewal rates remain stable")
        assert workbook["Operating_Drivers"]["H6"].value.startswith("Retention and implementation conversion")
        assert workbook["Quarter_Notes_UI"]["A10"].value == "Demand"
        assert workbook["Quarter_Notes_UI"]["C10"].value.startswith("Recurring software demand")
        assert workbook["Quarter_Notes_UI"]["H10"].value.startswith("Growth durability")
        assert workbook["Quarter_Notes_UI"]["M10"].value == source_ref
        assert workbook["Promise_Progress_UI"]["A61"].value == "Revenue"
        assert workbook["Promise_Progress_UI"]["C61"].value == "Revenue growth expected in the mid-single-digit range."
        assert workbook["Promise_Progress_UI"]["G61"].value == "Open"
        assert workbook["Promise_Progress_UI"]["I61"].value == "2026-Q1"
    finally:
        workbook.close()


def test_in_memory_execution_preserves_formulas_and_layout() -> None:
    template = load_workbook(BytesIO(TEMPLATE.read_bytes()), data_only=False, read_only=False)
    workbook, _plan, _written = _execute_in_memory()
    try:
        assert _formula_map(workbook) == _formula_map(template)
        assert _layout_signature(workbook) == _layout_signature(template)
    finally:
        template.close()
        workbook.close()


def test_failed_binding_plan_cannot_execute() -> None:
    package = deepcopy(_load_json(PACKAGE))
    package["company_profile"].pop("revenue_streams")
    failed_plan = _approved_plan(package)
    workbook = load_workbook(BytesIO(TEMPLATE.read_bytes()), data_only=False, read_only=False)
    try:
        with pytest.raises(BindingContractError, match="independently reproduced PASS"):
            _execute_binding_plan(workbook, failed_plan)
    finally:
        workbook.close()


def test_binding_semantic_drift_blocks_before_execution(tmp_path: Path) -> None:
    manifest = _load_json(MANIFEST)
    approved = _load_json(BINDING_MAP)
    identity = verify_shell_identity(TEMPLATE, manifest=manifest, binding_payload=approved)
    drifted = deepcopy(approved)
    next(row for row in drifted["bindings"] if row["binding_id"] == "summary_latest_revenue")["source_field"] = "net_income"

    plan = plan_standard_template_writes(
        _load_json(PACKAGE),
        binding_payload=drifted,
        manifest=manifest,
        shell_identity_report=identity,
    )

    assert plan.status == "FAIL"
    assert plan.planned_writes == []
    assert "shell_binding_contract_token_mismatch" in {issue.rule_id for issue in plan.issues}

    drifted_path = tmp_path / "drifted_bindings.json"
    drifted_path.write_text(json.dumps(drifted), encoding="utf-8")
    output_path = tmp_path / "must_not_exist.xlsx"
    with pytest.raises(NormalizedDataValidationError):
        fill_standard_template_from_package(PACKAGE, output_path=output_path, binding_map_path=drifted_path)
    assert not output_path.exists()


def test_filler_rejects_stale_expected_plan_before_copy(tmp_path: Path) -> None:
    package = _load_json(PACKAGE)
    manifest = _load_json(MANIFEST)
    bindings = _load_json(BINDING_MAP)
    identity = verify_shell_identity(TEMPLATE, manifest=manifest, binding_payload=bindings)
    plan, _snapshot = reproduce_binding_plan_snapshot(
        package,
        binding_payload=bindings,
        manifest=manifest,
        shell_path=TEMPLATE,
        shell_identity_report=identity,
    )
    stale = plan.to_dict()
    stale["planned_writes"][0]["value"] = "STALE EXPECTED VALUE"
    output_path = tmp_path / "must-not-be-copied.xlsx"

    with pytest.raises(BindingContractError, match="differs"):
        fill_standard_template_from_package(
            PACKAGE,
            output_path=output_path,
            expected_plan=stale,
        )

    assert not output_path.exists()


def test_missing_required_data_and_p1_semantics_block_before_execution(tmp_path: Path) -> None:
    missing = deepcopy(_load_json(PACKAGE))
    missing["company_profile"].pop("revenue_streams")
    missing_plan = _approved_plan(missing)

    invalid = deepcopy(_load_json(PACKAGE))
    invalid["normalized_guidance"]["items"][0]["value"]["value"] = "Adjusted EBITDA expected to improve by $10m."
    invalid_plan = _approved_plan(invalid)

    assert missing_plan.status == "FAIL"
    assert missing_plan.planned_writes == []
    assert invalid_plan.status == "FAIL"
    assert invalid_plan.planned_writes == []
    assert "guidance_metric_misclassification" in {issue.rule_id for issue in invalid_plan.issues}

    for name, package in (("missing", missing), ("invalid", invalid)):
        package_path = tmp_path / f"{name}.json"
        package_path.write_text(json.dumps(package), encoding="utf-8")
        output_path = tmp_path / f"{name}.xlsx"
        with pytest.raises(NormalizedDataValidationError):
            fill_standard_template_from_package(package_path, output_path=output_path)
        assert not output_path.exists()


def test_filler_contract_does_not_include_gtx_or_change_default_tickers() -> None:
    source = (ROOT / "pbi_xbrl" / "new_ticker_value_filler.py").read_text(encoding="utf-8")
    validation_runner = (ROOT / "pbi_xbrl" / "workbook_validation_runner.py").read_text(encoding="utf-8")

    assert "GTX" not in source
    assert 'TICKERS: Sequence[str] = ("PBI", "GPRE", "ANF")' in validation_runner
