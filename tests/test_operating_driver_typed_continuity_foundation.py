from __future__ import annotations

import dataclasses
import inspect
from datetime import date, datetime, timedelta

import pytest

from openpyxl import Workbook

from pbi_xbrl.excel_writer_operating_driver_workbook_support import (
    OperatingDriverWorkbookSupport,
    OperatingDriverWorkbookSupportDeps,
)
from pbi_xbrl.longitudinal_memory.calendar_rules import (
    CALENDAR_YEAR_RULE_ID,
    SOURCE_LABELLED_52_53_WEEK_RULE_ID,
)
from pbi_xbrl.longitudinal_memory.operating_driver_foundation import (
    AggregateCompleteness,
    AggregateReason,
    AggregationSemantics,
    ComparisonReason,
    ComparisonState,
    DefinitionContinuity,
    DefinitionContinuityState,
    DriverDimension,
    DriverIdentity,
    DurationAggregateRequest,
    DurationPeriod,
    EvidenceAvailability,
    EvidenceClassification,
    EvidenceSourceReference,
    EvidenceSourceType,
    EvidenceTransformation,
    EvidenceValueKind,
    FiscalCalendarIdentity,
    FiscalQuarterPeriod,
    FiscalYearPeriod,
    InstantPeriod,
    OperatingDriverEvidence,
    OperatingDriverFoundationError,
    PeriodKind,
    PeriodResolutionState,
    TrailingTwelveMonthsPeriod,
    UnitConversionReceipt,
    aggregate_duration_fail_closed,
    calendar_year_fiscal_year_period,
    calendar_year_quarter_period,
    foundation_record_sha256,
    prior_quarter_key,
    prior_year_quarter_key,
    resolve_exact_prior_fiscal_year,
    resolve_exact_prior_quarter,
    resolve_exact_prior_year_quarter,
    safe_qoq,
    safe_yoy,
    serialize_foundation_record,
    ttm_quarter_keys,
)
from pbi_xbrl.longitudinal_memory import operating_driver_foundation


def _date_or_none(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _gpre_45z_gap_workbook() -> Workbook:
    workbook = Workbook()
    drivers = workbook.active
    drivers.title = "Operating_Drivers"
    drivers.append(
        [
            "Quarter",
            "2025-Q1",
            "2025-Q2",
            "2025-Q3",
            "2025-Q4",
            "2026-Q1",
        ]
    )
    drivers.append(["45Z value realized ($m)", None, None, 26.0, 27.0, 55.2])

    history = workbook.create_sheet("History_Q")
    history.append(["quarter", "fiscal_year", "fiscal_quarter", "fiscal_label"])
    for fiscal_year, fiscal_quarter, end_date in (
        (2025, 1, date(2025, 3, 31)),
        (2025, 2, date(2025, 6, 30)),
        (2025, 3, date(2025, 9, 30)),
        (2025, 4, date(2025, 12, 31)),
        (2026, 1, date(2026, 3, 31)),
    ):
        history.append(
            [
                end_date,
                fiscal_year,
                fiscal_quarter,
                f"{fiscal_year}-Q{fiscal_quarter}",
            ]
        )
    return workbook


def test_gpre_45z_incomplete_fy_and_ttm_fail_closed() -> None:
    workbook = _gpre_45z_gap_workbook()
    support = OperatingDriverWorkbookSupport(
        OperatingDriverWorkbookSupportDeps(runtime={"_date_or_none": _date_or_none})
    )

    assert (
        support.operating_driver_latest_full_year_sum_from_workbook(
            workbook, "45Z value realized ($m)"
        )
        is None
    )
    assert (
        support.operating_driver_ttm_sum_from_workbook(
            workbook, "45Z value realized ($m)"
        )
        is None
    )

    fy_result = support.operating_driver_latest_full_year_result_from_workbook(
        workbook, "45Z value realized ($m)"
    )
    ttm_result = support.operating_driver_ttm_result_from_workbook(
        workbook, "45Z value realized ($m)"
    )
    assert fy_result is not None
    assert fy_result.completeness is AggregateCompleteness.INCOMPLETE
    assert fy_result.reason is AggregateReason.UNAVAILABLE_INCOMPLETE_PERIOD_SET
    assert fy_result.missing_constituent_period_ids == (
        "period:workbook:fy2025-q1@1",
        "period:workbook:fy2025-q2@1",
    )
    assert ttm_result is not None
    assert ttm_result.completeness is AggregateCompleteness.INCOMPLETE
    assert ttm_result.reason is AggregateReason.UNAVAILABLE_INCOMPLETE_PERIOD_SET
    assert ttm_result.missing_constituent_period_ids == (
        "period:workbook:fy2025-q2@1",
    )


def _calendar(company_id: str = "TEST") -> FiscalCalendarIdentity:
    slug = company_id.lower()
    return FiscalCalendarIdentity(
        calendar_id=f"calendar:{slug}:calendar-year@1",
        company_id=company_id,
        calendar_rule_id=CALENDAR_YEAR_RULE_ID,
        week_pattern="calendar",
    )


def _quarter(
    fiscal_year: int,
    fiscal_quarter: int,
    *,
    company_id: str = "TEST",
    calendar: FiscalCalendarIdentity | None = None,
) -> FiscalQuarterPeriod:
    calendar = calendar or _calendar(company_id)
    return calendar_year_quarter_period(
        company_id=company_id,
        calendar=calendar,
        fiscal_year=fiscal_year,
        fiscal_quarter=fiscal_quarter,
        period_id=f"period:{company_id.lower()}:fy{fiscal_year}-q{fiscal_quarter}@1",
    )


def _quarters(
    fiscal_year: int = 2025,
    *,
    company_id: str = "TEST",
) -> tuple[FiscalQuarterPeriod, ...]:
    calendar = _calendar(company_id)
    return tuple(
        _quarter(
            fiscal_year,
            fiscal_quarter,
            company_id=company_id,
            calendar=calendar,
        )
        for fiscal_quarter in range(1, 5)
    )


def _fiscal_year(
    fiscal_year: int = 2025,
    *,
    company_id: str = "TEST",
    calendar: FiscalCalendarIdentity | None = None,
) -> FiscalYearPeriod:
    calendar = calendar or _calendar(company_id)
    return calendar_year_fiscal_year_period(
        company_id=company_id,
        calendar=calendar,
        fiscal_year=fiscal_year,
        period_id=f"period:{company_id.lower()}:fy{fiscal_year}@1",
    )


def _dimension(
    member_id: str = "member:company:total@1", label: str = "Total company"
) -> DriverDimension:
    return DriverDimension(
        dimension_id="dimension:scope:company@1",
        member_id=member_id,
        label=label,
    )


def _driver(
    *,
    company_id: str = "TEST",
    driver_id: str = "driver:test:throughput@1",
    definition_id: str = "definition:test:throughput@1",
    unit_id: str = "unit:million-units@1",
    dimensions: tuple[DriverDimension, ...] | None = None,
    aggregation_semantics: AggregationSemantics = AggregationSemantics.SUMMABLE,
) -> DriverIdentity:
    return DriverIdentity(
        driver_id=driver_id,
        company_id=company_id,
        ticker=company_id,
        driver_family="throughput",
        canonical_label="Throughput",
        display_label="Throughput (m units)",
        unit_id=unit_id,
        scale="1000000",
        sign_convention="positive-is-more-throughput",
        dimensions=dimensions or (_dimension(),),
        period_kind=PeriodKind.FISCAL_QUARTER,
        source_owner="owner:operating-drivers:source-native@1",
        definition_id=definition_id,
        definition_version=1,
        aggregation_semantics=aggregation_semantics,
    )


def _same_series(driver: DriverIdentity) -> DefinitionContinuity:
    return DefinitionContinuity(
        state=DefinitionContinuityState.SAME_SERIES,
        from_definition_id=driver.definition_id,
        from_definition_version=driver.definition_version,
        to_definition_id=driver.definition_id,
        to_definition_version=driver.definition_version,
        reason="Same accepted definition.",
    )


def _source(
    *,
    source_type: EvidenceSourceType = EvidenceSourceType.EARNINGS_RELEASE,
    publication_date: date | None = date(2026, 2, 1),
    knowledge_date: date | None = date(2026, 2, 1),
) -> EvidenceSourceReference:
    return EvidenceSourceReference(
        source_document_id="doc:test:source",
        source_type=source_type,
        source_location="table:operating-kpis:row-1",
        publication_date=publication_date,
        knowledge_date=knowledge_date,
    )


def _evidence(
    driver: DriverIdentity,
    period: FiscalQuarterPeriod,
    value: str | None,
    *,
    evidence_id: str | None = None,
    availability: EvidenceAvailability | None = None,
    classification: EvidenceClassification = EvidenceClassification.ACTUAL,
    continuity: DefinitionContinuity | None = None,
    source_unit_id: str | None = None,
    raw_value: str | None = None,
) -> OperatingDriverEvidence:
    availability = availability or (
        EvidenceAvailability.AVAILABLE
        if value is not None
        else EvidenceAvailability.UNAVAILABLE
    )
    return OperatingDriverEvidence(
        evidence_id=evidence_id or f"evidence:{driver.company_id.lower()}:{period.period_id}",
        driver=driver,
        period=period,
        source=_source(),
        value_kind=EvidenceValueKind.NUMERIC,
        raw_value=(value if raw_value is None else raw_value),
        normalized_value=value,
        source_unit_id=source_unit_id or driver.unit_id,
        classification=classification,
        availability=availability,
        unavailable_reason=(
            None
            if availability is EvidenceAvailability.AVAILABLE
            else "MISSING_SOURCE_VALUE"
        ),
        continuity=continuity or _same_series(driver),
    )


def _fy_request(
    driver: DriverIdentity,
    periods: tuple[FiscalQuarterPeriod, ...],
) -> DurationAggregateRequest:
    return DurationAggregateRequest(
        request_id=f"aggregate:{driver.company_id.lower()}:throughput:fy",
        driver=driver,
        requested_period=_fiscal_year(
            periods[0].fiscal_year,
            company_id=driver.company_id,
            calendar=periods[0].calendar,
        ),
        required_constituent_quarters=periods,
    )


def _ttm_request(
    driver: DriverIdentity,
    periods: tuple[FiscalQuarterPeriod, ...],
) -> DurationAggregateRequest:
    return DurationAggregateRequest(
        request_id=f"aggregate:{driver.company_id.lower()}:throughput:ttm",
        driver=driver,
        requested_period=TrailingTwelveMonthsPeriod(
            period_id=(
                f"period:{driver.company_id.lower()}:ttm-"
                f"fy{periods[-1].fiscal_year}-q{periods[-1].fiscal_quarter}@1"
            ),
            company_id=driver.company_id,
            ending_quarter=periods[-1],
            constituent_quarters=periods,
        ),
        required_constituent_quarters=periods,
    )


def test_source_backed_all_zero_workbook_aggregate_remains_numeric_zero() -> None:
    workbook = _gpre_45z_gap_workbook()
    drivers = workbook["Operating_Drivers"]
    for column in range(2, 7):
        drivers.cell(2, column, 0.0)
    support = OperatingDriverWorkbookSupport(
        OperatingDriverWorkbookSupportDeps(runtime={"_date_or_none": _date_or_none})
    )

    assert support.operating_driver_latest_full_year_sum_from_workbook(
        workbook, "45Z value realized ($m)"
    ) == pytest.approx(0.0)
    assert support.operating_driver_ttm_sum_from_workbook(
        workbook, "45Z value realized ($m)"
    ) == pytest.approx(0.0)


def test_complete_four_quarter_sum_is_available() -> None:
    periods = _quarters()
    driver = _driver()
    result = aggregate_duration_fail_closed(
        _fy_request(driver, periods),
        tuple(
            _evidence(driver, period, value)
            for period, value in zip(periods, ("1", "2", "3", "4"))
        ),
    )

    assert result.completeness is AggregateCompleteness.COMPLETE
    assert result.reason is AggregateReason.COMPLETE
    assert result.result_available is True
    assert result.value == "10"


def test_one_missing_quarter_emits_no_numeric_aggregate() -> None:
    periods = _quarters()
    driver = _driver()
    observations = tuple(
        _evidence(driver, period, None if index == 1 else str(index + 1))
        for index, period in enumerate(periods)
    )

    result = aggregate_duration_fail_closed(_fy_request(driver, periods), observations)

    assert result.completeness is AggregateCompleteness.INCOMPLETE
    assert result.reason is AggregateReason.UNAVAILABLE_INCOMPLETE_PERIOD_SET
    assert result.value is None
    assert result.missing_constituent_period_ids == (periods[1].period_id,)


def test_explicit_source_backed_zero_is_a_valid_constituent() -> None:
    periods = _quarters()
    driver = _driver()
    zero = _evidence(driver, periods[1], "0")
    observations = (
        _evidence(driver, periods[0], "1"),
        zero,
        _evidence(driver, periods[2], "2"),
        _evidence(driver, periods[3], "3"),
    )

    result = aggregate_duration_fail_closed(_fy_request(driver, periods), observations)

    assert zero.source_backed_zero is True
    assert result.completeness is AggregateCompleteness.COMPLETE
    assert result.value == "6"


def test_duplicate_constituent_period_is_rejected() -> None:
    periods = _quarters()
    driver = _driver()
    observations = [
        _evidence(driver, period, str(index + 1))
        for index, period in enumerate(periods)
    ]
    observations.append(
        dataclasses.replace(observations[0], evidence_id="evidence:test:duplicate")
    )

    result = aggregate_duration_fail_closed(_fy_request(driver, periods), observations)

    assert result.completeness is AggregateCompleteness.INCOMPATIBLE
    assert result.reason is AggregateReason.DUPLICATE_CONSTITUENT_PERIOD
    assert result.duplicate_constituent_period_ids == (periods[0].period_id,)


def test_wrong_fiscal_quarter_is_rejected_not_window_shrunk() -> None:
    periods = _quarters()
    driver = _driver()
    wrong_period = _quarter(2024, 4)
    observations = [
        _evidence(driver, period, str(index + 1))
        for index, period in enumerate(periods[1:])
    ]
    observations.append(_evidence(driver, wrong_period, "1"))

    result = aggregate_duration_fail_closed(_fy_request(driver, periods), observations)

    assert result.completeness is AggregateCompleteness.INCOMPATIBLE
    assert result.reason is AggregateReason.UNEXPECTED_CONSTITUENT_PERIOD
    assert result.unexpected_constituent_period_ids == (wrong_period.period_id,)


def test_definition_break_inside_aggregate_is_rejected() -> None:
    periods = _quarters()
    driver = _driver()
    observations = [
        _evidence(driver, period, str(index + 1))
        for index, period in enumerate(periods)
    ]
    break_continuity = DefinitionContinuity(
        state=DefinitionContinuityState.DEFINITION_CHANGED_BREAK_SERIES,
        from_definition_id="definition:test:legacy-throughput@1",
        from_definition_version=1,
        to_definition_id=driver.definition_id,
        to_definition_version=driver.definition_version,
        reason="Issuer changed the throughput definition.",
    )
    observations[2] = dataclasses.replace(
        observations[2], continuity=break_continuity
    )

    result = aggregate_duration_fail_closed(_fy_request(driver, periods), observations)

    assert result.completeness is AggregateCompleteness.INCOMPATIBLE
    assert result.reason is AggregateReason.DEFINITION_INCOMPATIBLE


def test_dimension_change_inside_aggregate_is_rejected() -> None:
    periods = _quarters()
    driver = _driver()
    brand_driver = _driver(
        dimensions=(
            DriverDimension(
                dimension_id="dimension:brand:brand@1",
                member_id="member:brand:brand-a@1",
                label="Brand A",
            ),
        )
    )
    observations = [
        _evidence(driver, period, str(index + 1))
        for index, period in enumerate(periods)
    ]
    observations[1] = _evidence(brand_driver, periods[1], "2")

    result = aggregate_duration_fail_closed(_fy_request(driver, periods), observations)

    assert result.completeness is AggregateCompleteness.INCOMPATIBLE
    assert result.reason is AggregateReason.DIMENSION_INCOMPATIBLE


def test_incompatible_unit_inside_aggregate_is_rejected() -> None:
    periods = _quarters()
    driver = _driver()
    barrel_driver = _driver(unit_id="unit:barrels@1")
    observations = [
        _evidence(driver, period, str(index + 1))
        for index, period in enumerate(periods)
    ]
    observations[1] = _evidence(barrel_driver, periods[1], "2")

    result = aggregate_duration_fail_closed(_fy_request(driver, periods), observations)

    assert result.completeness is AggregateCompleteness.INCOMPATIBLE
    assert result.reason is AggregateReason.UNIT_INCOMPATIBLE


@pytest.mark.parametrize(
    "aggregation_semantics",
    [
        AggregationSemantics.PERIOD_END,
        AggregationSemantics.AVERAGE_REQUIRES_CONTRACT,
        AggregationSemantics.NON_AGGREGATABLE,
        AggregationSemantics.UNKNOWN,
    ],
)
def test_non_summable_metric_semantics_reject_sum(
    aggregation_semantics: AggregationSemantics,
) -> None:
    periods = _quarters()
    driver = _driver(aggregation_semantics=aggregation_semantics)

    result = aggregate_duration_fail_closed(
        _fy_request(driver, periods),
        tuple(_evidence(driver, period, "1") for period in periods),
    )

    assert result.completeness is AggregateCompleteness.INCOMPATIBLE
    assert result.reason is AggregateReason.AGGREGATION_SEMANTICS_INVALID


def test_typed_fiscal_quarter_is_not_its_display_label() -> None:
    period = _quarter(2026, 1, company_id="PBI")

    assert period.period_kind is PeriodKind.FISCAL_QUARTER
    assert period.display_label == "2026-Q1"
    assert period.period_id == "period:pbi:fy2026-q1@1"
    assert period.start_date == date(2026, 1, 1)
    assert period.end_date == date(2026, 3, 31)
    assert period.fiscal_ordinal == 2026 * 4


def test_typed_fiscal_year_has_exact_duration_identity() -> None:
    period = _fiscal_year(2025, company_id="GPRE")

    assert period.period_kind is PeriodKind.FISCAL_YEAR
    assert period.start_date == date(2025, 1, 1)
    assert period.end_date == date(2025, 12, 31)
    assert period.day_count == 365
    assert period.display_label == "2025"


def test_typed_ttm_resolves_exact_four_constituents() -> None:
    periods = (
        _quarter(2025, 2),
        _quarter(2025, 3),
        _quarter(2025, 4),
        _quarter(2026, 1),
    )
    ttm = TrailingTwelveMonthsPeriod(
        period_id="period:test:ttm-fy2026-q1@1",
        company_id="TEST",
        ending_quarter=periods[-1],
        constituent_quarters=periods,
    )

    assert ttm.period_kind is PeriodKind.TTM
    assert tuple(
        (period.fiscal_year, period.fiscal_quarter)
        for period in ttm.constituent_quarters
    ) == ttm_quarter_keys(2026, 1)
    assert ttm.display_label == "TTM through 2026-Q1"


def test_typed_ttm_rejects_a_disclosure_gap() -> None:
    periods = (
        _quarter(2025, 1),
        _quarter(2025, 2),
        _quarter(2025, 4),
        _quarter(2026, 1),
    )

    with pytest.raises(
        OperatingDriverFoundationError,
        match="exact, adjacent fiscal quarters",
    ):
        TrailingTwelveMonthsPeriod(
            period_id="period:test:ttm-fy2026-q1@1",
            company_id="TEST",
            ending_quarter=periods[-1],
            constituent_quarters=periods,
        )


def test_instant_and_duration_are_distinct_period_contracts() -> None:
    instant = InstantPeriod(
        period_id="period:test:instant-2026-03-31@1",
        company_id="TEST",
        instant_date=date(2026, 3, 31),
        fiscal_calendar_id=_calendar().calendar_id,
    )
    duration = DurationPeriod(
        period_id="period:test:duration-2026-q1@1",
        company_id="TEST",
        start_date=date(2026, 1, 1),
        end_date=date(2026, 3, 31),
        fiscal_calendar_id=_calendar().calendar_id,
    )

    assert instant.period_kind is PeriodKind.INSTANT
    assert duration.period_kind is PeriodKind.DURATION
    assert "instant_date" in instant.to_dict()
    assert "start_date" in duration.to_dict()


def _anf_13_week_quarter(
    fiscal_year: int,
    fiscal_quarter: int,
    start_date: date,
    *,
    fiscal_ordinal: int,
) -> FiscalQuarterPeriod:
    calendar = FiscalCalendarIdentity(
        calendar_id="calendar:anf:source-labelled-fiscal@1",
        company_id="ANF",
        calendar_rule_id=SOURCE_LABELLED_52_53_WEEK_RULE_ID,
        week_pattern="4-4-5",
    )
    return FiscalQuarterPeriod(
        period_id=f"period:anf:fy{fiscal_year}-q{fiscal_quarter}@1",
        company_id="ANF",
        calendar=calendar,
        fiscal_year=fiscal_year,
        fiscal_quarter=fiscal_quarter,
        fiscal_ordinal=fiscal_ordinal,
        start_date=start_date,
        end_date=start_date + timedelta(days=90),
        week_count=13,
        is_53_week_year=False,
    )


def test_non_calendar_fiscal_identity_is_supported_without_calendar_labels() -> None:
    q1 = _anf_13_week_quarter(
        2026, 1, date(2025, 2, 2), fiscal_ordinal=105
    )
    q2 = _anf_13_week_quarter(
        2026, 2, date(2025, 5, 4), fiscal_ordinal=106
    )

    resolution = resolve_exact_prior_quarter(q2, (q1,))

    assert resolution.state is PeriodResolutionState.RESOLVED
    assert resolution.period == q1
    assert q1.end_date == date(2025, 5, 3)
    assert q1.display_label == "2026-Q1"


def test_exact_prior_quarter_resolves_across_fiscal_year_boundary() -> None:
    current = _quarter(2026, 1)
    prior = _quarter(2025, 4)

    resolution = resolve_exact_prior_quarter(current, (prior, _quarter(2025, 3)))

    assert prior_quarter_key(2026, 1) == (2025, 4)
    assert resolution.state is PeriodResolutionState.RESOLVED
    assert resolution.period == prior


def test_qoq_gap_does_not_bridge_to_prior_available_quarter() -> None:
    current = _quarter(2025, 4)
    q2 = _quarter(2025, 2)

    resolution = resolve_exact_prior_quarter(current, (q2,))

    assert resolution.state is PeriodResolutionState.PRIOR_PERIOD_MISSING
    assert resolution.period is None
    assert resolution.expected_fiscal_quarter == 3


def test_exact_prior_year_quarter_resolves() -> None:
    current = _quarter(2026, 2)
    prior = _quarter(2025, 2)

    resolution = resolve_exact_prior_year_quarter(current, (prior,))

    assert prior_year_quarter_key(2026, 2) == (2025, 2)
    assert resolution.state is PeriodResolutionState.RESOLVED
    assert resolution.period == prior


def test_yoy_does_not_use_wrong_prior_year_quarter() -> None:
    current = _quarter(2026, 2)

    resolution = resolve_exact_prior_year_quarter(current, (_quarter(2025, 1),))

    assert resolution.state is PeriodResolutionState.PRIOR_YEAR_PERIOD_MISSING
    assert resolution.period is None


def test_exact_prior_fiscal_year_resolves_by_typed_identity() -> None:
    current = _fiscal_year(2026)
    prior = _fiscal_year(2025)

    resolution = resolve_exact_prior_fiscal_year(current, (prior,))

    assert resolution.state is PeriodResolutionState.RESOLVED
    assert resolution.period == prior


def test_driver_identity_includes_dimensions_not_workbook_coordinates() -> None:
    driver = _driver()
    payload = driver.to_dict()

    assert payload["driver_id"] == "driver:test:throughput@1"
    assert payload["dimension_set_id"].startswith("dimset:v1|")
    assert "row" not in payload
    assert "cell" not in payload


@pytest.mark.parametrize(
    ("state", "automatic_join_safe"),
    [
        (DefinitionContinuityState.SAME_SERIES, True),
        (DefinitionContinuityState.RESTATED_SAME_SERIES, True),
        (DefinitionContinuityState.DEFINITION_CHANGED_BREAK_SERIES, False),
        (DefinitionContinuityState.UNIT_CONVERSION_SAFE, True),
        (DefinitionContinuityState.SEGMENT_REORG_BREAK_SERIES, False),
        (DefinitionContinuityState.SUCCESSOR_METRIC, False),
        (DefinitionContinuityState.UNRESOLVED, False),
    ],
)
def test_all_definition_continuity_states_are_typed(
    state: DefinitionContinuityState, automatic_join_safe: bool
) -> None:
    driver = _driver()
    kwargs: dict[str, object] = {}
    from_definition = driver.definition_id
    if state is DefinitionContinuityState.RESTATED_SAME_SERIES:
        kwargs["restatement_id"] = "restatement:test:throughput-fy2025@1"
    if state is DefinitionContinuityState.UNIT_CONVERSION_SAFE:
        kwargs["unit_conversion"] = UnitConversionReceipt(
            rule_id="rule:operating-drivers:thousand-to-million@1",
            from_unit_id="unit:thousand-units@1",
            to_unit_id=driver.unit_id,
            multiplier="0.001",
            from_scale="1000",
            to_scale="1000000",
        )
    if state is DefinitionContinuityState.SUCCESSOR_METRIC:
        kwargs["successor_driver_id"] = "driver:test:throughput-successor@1"
    if state is DefinitionContinuityState.DEFINITION_CHANGED_BREAK_SERIES:
        from_definition = "definition:test:legacy-throughput@1"
    continuity = DefinitionContinuity(
        state=state,
        from_definition_id=from_definition,
        from_definition_version=1,
        to_definition_id=driver.definition_id,
        to_definition_version=driver.definition_version,
        reason=f"Representative {state.value} fixture.",
        **kwargs,
    )

    assert continuity.state is state
    assert continuity.automatic_join_safe is automatic_join_safe


def test_same_series_cannot_mask_a_definition_change() -> None:
    driver = _driver()

    with pytest.raises(OperatingDriverFoundationError, match="SAME_SERIES"):
        DefinitionContinuity(
            state=DefinitionContinuityState.SAME_SERIES,
            from_definition_id="definition:test:legacy-throughput@1",
            from_definition_version=1,
            to_definition_id=driver.definition_id,
            to_definition_version=driver.definition_version,
            reason="Incorrectly claimed continuity.",
        )


def test_safe_unit_conversion_has_a_deterministic_receipt() -> None:
    driver = _driver()
    receipt = UnitConversionReceipt(
        rule_id="rule:operating-drivers:thousand-to-million@1",
        from_unit_id="unit:thousand-units@1",
        to_unit_id=driver.unit_id,
        multiplier="0.001",
        from_scale="1000",
        to_scale="1000000",
    )
    continuity = DefinitionContinuity(
        state=DefinitionContinuityState.UNIT_CONVERSION_SAFE,
        from_definition_id=driver.definition_id,
        from_definition_version=1,
        to_definition_id=driver.definition_id,
        to_definition_version=1,
        reason="Only the disclosed scale changed.",
        unit_conversion=receipt,
    )
    evidence = _evidence(
        driver,
        _quarter(2025, 1),
        "1.25",
        raw_value="1250",
        source_unit_id="unit:thousand-units@1",
        continuity=continuity,
    )

    assert receipt.convert("1250") == "1.25"
    assert evidence.normalized_value == "1.25"
    assert evidence.continuity.state is DefinitionContinuityState.UNIT_CONVERSION_SAFE


def test_unsafe_unit_change_is_rejected_at_evidence_boundary() -> None:
    driver = _driver()

    with pytest.raises(
        OperatingDriverFoundationError,
        match="safe conversion receipt",
    ):
        _evidence(
            driver,
            _quarter(2025, 1),
            "1.25",
            raw_value="1250",
            source_unit_id="unit:thousand-units@1",
        )


def test_typed_lineage_preserves_source_and_knowledge_date() -> None:
    driver = _driver(company_id="GPRE")
    period = _quarter(2026, 1, company_id="GPRE")
    evidence = OperatingDriverEvidence(
        evidence_id="evidence:gpre:production:fy2026-q1@1",
        driver=driver,
        period=period,
        source=EvidenceSourceReference(
            source_document_id="doc:gpre:q1-2026-release",
            source_type=EvidenceSourceType.EARNINGS_RELEASE,
            source_location="table:operating-data:ethanol-production",
            publication_date=date(2026, 5, 7),
            knowledge_date=date(2026, 5, 7),
        ),
        value_kind=EvidenceValueKind.NUMERIC,
        raw_value="248.4",
        normalized_value="248.4",
        source_unit_id=driver.unit_id,
        classification=EvidenceClassification.ACTUAL,
        availability=EvidenceAvailability.AVAILABLE,
        unavailable_reason=None,
        continuity=_same_series(driver),
        transformations=(
            EvidenceTransformation(
                method_id="method:operating-drivers:source-normalization@1",
                description="Preserved the source value in the canonical unit.",
                input_record_ids=("occ:gpre:q1-2026:production",),
            ),
        ),
    )
    payload = evidence.to_dict()

    assert payload["source"]["publication_date"] == "2026-05-07"
    assert payload["source"]["knowledge_date"] == "2026-05-07"
    assert payload["source"]["source_location"].startswith("table:")
    assert payload["period"]["period_id"] == period.period_id
    assert payload["driver"]["definition_id"] == driver.definition_id


def test_unknown_knowledge_date_remains_unknown() -> None:
    source = _source(publication_date=None, knowledge_date=None)

    assert source.to_dict()["publication_date"] is None
    assert source.to_dict()["knowledge_date"] is None


def test_knowledge_date_cannot_precede_publication() -> None:
    with pytest.raises(OperatingDriverFoundationError, match="cannot precede"):
        _source(
            publication_date=date(2026, 5, 7),
            knowledge_date=date(2026, 5, 6),
        )


def test_safe_qoq_uses_only_exact_prior_actual() -> None:
    driver = _driver()
    prior = _evidence(driver, _quarter(2025, 1), "10")
    current = _evidence(driver, _quarter(2025, 2), "12")

    result = safe_qoq(current, (prior,))

    assert result.state is ComparisonState.COMPLETE
    assert result.reason is ComparisonReason.COMPLETE
    assert result.absolute_change == "2"
    assert result.percent_change == "0.2"


def test_safe_qoq_returns_prior_period_missing_instead_of_bridging() -> None:
    driver = _driver()
    q2 = _evidence(driver, _quarter(2025, 2), "10")
    q4 = _evidence(driver, _quarter(2025, 4), "12")

    result = safe_qoq(q4, (q2,))

    assert result.state is ComparisonState.UNAVAILABLE
    assert result.reason is ComparisonReason.PRIOR_PERIOD_MISSING
    assert result.absolute_change is None


def test_safe_yoy_uses_only_exact_same_fiscal_quarter() -> None:
    driver = _driver()
    prior = _evidence(driver, _quarter(2025, 2), "10")
    current = _evidence(driver, _quarter(2026, 2), "15")

    result = safe_yoy(current, (prior,))

    assert result.state is ComparisonState.COMPLETE
    assert result.absolute_change == "5"
    assert result.percent_change == "0.5"


def test_safe_yoy_returns_prior_year_period_missing() -> None:
    driver = _driver()
    wrong_prior = _evidence(driver, _quarter(2025, 1), "10")
    current = _evidence(driver, _quarter(2026, 2), "15")

    result = safe_yoy(current, (wrong_prior,))

    assert result.state is ComparisonState.UNAVAILABLE
    assert result.reason is ComparisonReason.PRIOR_YEAR_PERIOD_MISSING


def test_safe_comparison_rejects_definition_break() -> None:
    driver = _driver()
    prior = _evidence(driver, _quarter(2025, 1), "10")
    broken = DefinitionContinuity(
        state=DefinitionContinuityState.DEFINITION_CHANGED_BREAK_SERIES,
        from_definition_id="definition:test:legacy-throughput@1",
        from_definition_version=1,
        to_definition_id=driver.definition_id,
        to_definition_version=1,
        reason="Definition changed at Q2.",
    )
    current = _evidence(
        driver, _quarter(2025, 2), "12", continuity=broken
    )

    result = safe_qoq(current, (prior,))

    assert result.state is ComparisonState.INCOMPATIBLE
    assert result.reason is ComparisonReason.DEFINITION_INCOMPATIBLE


def test_safe_comparison_rejects_guidance_to_actual_by_default() -> None:
    driver = _driver()
    prior = _evidence(driver, _quarter(2025, 1), "10")
    current = _evidence(
        driver,
        _quarter(2025, 2),
        "12",
        classification=EvidenceClassification.GUIDANCE,
    )

    result = safe_qoq(current, (prior,))

    assert result.state is ComparisonState.INCOMPATIBLE
    assert result.reason is ComparisonReason.CLASSIFICATION_INCOMPATIBLE


def test_zero_base_keeps_absolute_change_but_not_percent_change() -> None:
    driver = _driver()
    prior = _evidence(driver, _quarter(2025, 1), "0")
    current = _evidence(driver, _quarter(2025, 2), "2")

    result = safe_qoq(current, (prior,))

    assert result.state is ComparisonState.COMPLETE
    assert result.absolute_change == "2"
    assert result.percent_change is None
    assert result.percent_change_reason == "ZERO_BASE_PERCENT_CHANGE_UNAVAILABLE"


def test_pbi_sparse_qualitative_evidence_is_truthful_not_numeric() -> None:
    driver = _driver(
        company_id="PBI",
        driver_id="driver:pbi:presort-execution@1",
        definition_id="definition:pbi:presort-execution@1",
        unit_id="unit:text@1",
        aggregation_semantics=AggregationSemantics.NON_AGGREGATABLE,
    )
    period = _quarter(2026, 2, company_id="PBI")
    evidence = OperatingDriverEvidence(
        evidence_id="evidence:pbi:presort:fy2026-q2@1",
        driver=driver,
        period=period,
        source=_source(source_type=EvidenceSourceType.TRANSCRIPT),
        value_kind=EvidenceValueKind.QUALITATIVE,
        raw_value="Presort execution remained a management focus.",
        normalized_value=None,
        source_unit_id=driver.unit_id,
        classification=EvidenceClassification.ACTUAL,
        availability=EvidenceAvailability.AVAILABLE,
        unavailable_reason=None,
        continuity=_same_series(driver),
    )

    assert evidence.value_kind is EvidenceValueKind.QUALITATIVE
    assert evidence.normalized_value is None
    assert evidence.source_backed_zero is False


def test_pbi_missing_numeric_history_remains_unavailable() -> None:
    driver = _driver(company_id="PBI")
    evidence = _evidence(driver, _quarter(2026, 2, company_id="PBI"), None)

    assert evidence.availability is EvidenceAvailability.UNAVAILABLE
    assert evidence.normalized_value is None
    assert evidence.source_backed_zero is False


def test_gpre_production_is_summable_but_utilization_is_not() -> None:
    periods = _quarters(company_id="GPRE")
    production = _driver(
        company_id="GPRE",
        driver_id="driver:gpre:ethanol-production-volume@1",
        definition_id="definition:gpre:ethanol-production-volume@1",
        unit_id="unit:million-gallons@1",
        aggregation_semantics=AggregationSemantics.SUMMABLE,
    )
    utilization = _driver(
        company_id="GPRE",
        driver_id="driver:gpre:utilization@1",
        definition_id="definition:gpre:utilization@1",
        unit_id="unit:ratio@1",
        aggregation_semantics=AggregationSemantics.AVERAGE_REQUIRES_CONTRACT,
    )

    production_result = aggregate_duration_fail_closed(
        _fy_request(production, periods),
        tuple(_evidence(production, period, "100") for period in periods),
    )
    utilization_result = aggregate_duration_fail_closed(
        _fy_request(utilization, periods),
        tuple(_evidence(utilization, period, "0.9") for period in periods),
    )

    assert production_result.value == "400"
    assert utilization_result.reason is AggregateReason.AGGREGATION_SEMANTICS_INVALID


def test_gpre_45z_and_crush_definitions_do_not_form_one_series() -> None:
    prior_driver = _driver(
        company_id="GPRE",
        driver_id="driver:gpre:crush-margin@1",
        definition_id="definition:gpre:crush-margin@1",
        unit_id="unit:usd-million@1",
    )
    current_driver = _driver(
        company_id="GPRE",
        driver_id="driver:gpre:45z-value@1",
        definition_id="definition:gpre:45z-value@1",
        unit_id="unit:usd-million@1",
    )
    prior = _evidence(prior_driver, _quarter(2025, 1, company_id="GPRE"), "10")
    current = _evidence(current_driver, _quarter(2025, 2, company_id="GPRE"), "12")

    result = safe_qoq(current, (prior,))

    assert result.state is ComparisonState.INCOMPATIBLE
    assert result.reason is ComparisonReason.DEFINITION_INCOMPATIBLE


def test_deterministic_serialization_has_stable_hash_and_no_object_repr() -> None:
    record = _evidence(_driver(), _quarter(2025, 1), "12.50")

    first = serialize_foundation_record(record)
    second = serialize_foundation_record(record)

    assert first == second
    assert foundation_record_sha256(record) == foundation_record_sha256(record)
    assert b"object at 0x" not in first
    assert b'"normalized_value": "12.5"' in first


def test_dimension_input_order_does_not_change_driver_identity_serialization() -> None:
    first_dimension = DriverDimension(
        dimension_id="dimension:geography:region@1",
        member_id="member:geography:americas@1",
        label="Americas",
    )
    second_dimension = DriverDimension(
        dimension_id="dimension:product:family@1",
        member_id="member:product:core@1",
        label="Core",
    )
    first = _driver(dimensions=(first_dimension, second_dimension))
    second = _driver(dimensions=(second_dimension, first_dimension))

    assert first.dimension_set_id == second.dimension_set_id
    assert serialize_foundation_record(first) == serialize_foundation_record(second)


def test_non_anf_ticker_uses_the_same_typed_contract() -> None:
    driver = _driver(company_id="PBI")
    period = _quarter(2026, 2, company_id="PBI")
    evidence = _evidence(driver, period, "7")

    assert driver.company_id == "PBI"
    assert evidence.period.company_id == "PBI"
    assert "ANF" not in serialize_foundation_record(evidence).decode("utf-8")


def test_legacy_adapter_fails_closed_for_untyped_noncalendar_history() -> None:
    workbook = _gpre_45z_gap_workbook()
    history = workbook["History_Q"]
    history.cell(2, 1, date(2025, 5, 3))
    support = OperatingDriverWorkbookSupport(
        OperatingDriverWorkbookSupportDeps(runtime={"_date_or_none": _date_or_none})
    )

    assert support.operating_driver_latest_full_year_sum_from_workbook(
        workbook, "45Z value realized ($m)"
    ) is None
    assert support.operating_driver_ttm_sum_from_workbook(
        workbook, "45Z value realized ($m)"
    ) is None


def test_foundation_contains_no_ticker_specific_economic_branch() -> None:
    source = inspect.getsource(operating_driver_foundation)

    assert 'ticker == "ANF"' not in source
    assert 'ticker == "GPRE"' not in source
    assert 'ticker == "PBI"' not in source
    assert "45Z" not in source
