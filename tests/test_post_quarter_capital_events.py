from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from pbi_xbrl.excel_writer_post_quarter_capital_events import (
    write_post_quarter_capital_events_sheet,
)
from pbi_xbrl.post_quarter_capital_events import (
    apply_pbi_current_debt_overlay,
    build_post_quarter_capital_events,
)


def _write(path: Path, text: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
    return path


def _write_refresh_log(
    path: Path,
    *,
    ticker: str,
    form: str,
    filing_date: str,
    accession: str,
    destinations: list[Path],
) -> Path:
    payload = {
        "downloaded_at": "2026-06-25T19:49:52.168707+00:00",
        "records": [
            {
                "ticker": ticker,
                "form": form,
                "filing_date": filing_date,
                "accession": accession,
                "official_filename": destination.name,
                "destination": str(destination),
                "status": "downloaded",
            }
            for destination in destinations
        ],
    }
    path.write_text(json.dumps(payload), encoding="utf-8")
    return path


def test_pbi_parser_extracts_refinancing_event_without_numeric_false_positives(
    tmp_path: Path,
) -> None:
    material_root = tmp_path / "tickers" / "PBI"
    filing = _write(
        material_root / "financial_statement" / "d88573d8k.htm",
        """
        <html><body>
        <p>On June 23, 2026, the Company completed the redemption of all
        $347 million aggregate principal amount of its 6.875% Senior Notes due
        March 2027.</p>
        </body></html>
        """,
    )
    amendment = _write(
        material_root / "financial_statement" / "d88573dex101.htm",
        """
        <html><body>
        <p>The incremental Term Loan A commitment is $150 million.</p>
        <p>Following the amendment, total Term Loan A borrowings are $302 million.</p>
        <p>The Term Loan A maturity remains May 18, 2031.</p>
        </body></html>
        """,
    )
    release = _write(
        material_root / "press_release" / "d88573dex991.htm",
        """
        <html><body>
        <p>The transaction removes the March 2027 notes. The next scheduled
        debt maturity is March 2029. Existing cash and other liquidity were
        used with the incremental term loan to fund redemption and fees,
        costs and expenses.</p>
        </body></html>
        """,
    )
    refresh_log = _write_refresh_log(
        tmp_path / "source_refresh.json",
        ticker="PBI",
        form="8-K",
        filing_date="2026-06-25",
        accession="000119312526281893",
        destinations=[filing, amendment, release],
    )

    events = build_post_quarter_capital_events(
        ticker="PBI",
        material_roots=[material_root],
        cache_roots=[],
        source_refresh_logs=[refresh_log],
    )

    assert len(events) == 1
    event = events.iloc[0]
    assert event["event_type"] == "refinancing_redemption"
    assert event["reported_quarter_anchor"] == "2026-Q1"
    assert event["filing_type"] == "8-K package / exhibits"
    assert event["filing_date"] == "2026-06-25"
    assert event["downloaded_at"] == "2026-06-25T19:49:52.168707+00:00"
    assert event["accession"] == "000119312526281893"
    assert event["principal_redeemed"] == 347_000_000.0
    assert event["incremental_term_loan"] == 150_000_000.0
    assert event["term_loan_total"] == 302_000_000.0
    assert event["gross_principal_delta"] == -197_000_000.0
    assert event["next_scheduled_maturity"] == "March 2029"
    assert event["term_loan_maturity"] == "2031-05-18"
    assert event["automatic_net_debt_adjustment"] is False
    assert event["used_in_workbook"] == "Yes"
    assert "Valuation current Debt Detail" in event["used_surfaces"]
    assert event["source_path_exists"] is True
    assert "6.875" not in {
        event["principal_redeemed"],
        event["incremental_term_loan"],
        event["term_loan_total"],
    }
    assert 2027 not in {
        event["principal_redeemed"],
        event["incremental_term_loan"],
        event["term_loan_total"],
    }


def test_gpre_parser_keeps_warrant_count_and_max_issuable_shares_distinct(
    tmp_path: Path,
) -> None:
    material_root = tmp_path / "tickers" / "GPRE"
    purchase = _write(
        material_root / "financial_statement" / "tm2618355d1_ex10-2.htm",
        """
        <html><body>
        <p>The Company shall issue warrants to purchase 500,000 shares of
        common stock.</p>
        </body></html>
        """,
    )
    warrant_amounts = (366_240, 37_120, 10_360, 86_280)
    warrant_suffixes = ("a", "b", "c", "d")
    warrants = [
        _write(
            material_root
            / "financial_statement"
            / f"tm2618355d2_ex4-4{suffix}.htm",
            f"""
            <html><body>
            <p>Number of Warrants: {amount:,}</p>
            <p>Each Warrant Share may be purchased for a price of $0.01 per share,
            the Per Share Exercise Price.</p>
            <p>Before 5:00 P.M. on June 16, 2036 (the Expiration Date).</p>
            <p>The Beneficial Ownership Limitation shall be 19.8% of the number
            of shares of Common Stock outstanding.</p>
            </body></html>
            """,
        )
        for suffix, amount in zip(warrant_suffixes, warrant_amounts)
    ]
    prospectus = _write(
        material_root / "financial_statement" / "tm2618355d1_s3asr.htm",
        """
        <html><body>
        <p>This prospectus relates to the offer and sale by the selling
        stockholders of up to 550,000 shares of common stock issuable on the
        exercise of outstanding warrants.</p>
        <p>The amount represents the maximum number of shares issuable upon
        exercise without regard to the beneficial ownership limitation.</p>
        </body></html>
        """,
    )
    refresh_log = _write_refresh_log(
        tmp_path / "source_refresh.json",
        ticker="GPRE",
        form="S-3ASR",
        filing_date="2026-06-22",
        accession="000110465926076397",
        destinations=[purchase, *warrants, prospectus],
    )

    events = build_post_quarter_capital_events(
        ticker="GPRE",
        material_roots=[material_root],
        cache_roots=[],
        source_refresh_logs=[refresh_log],
    )

    assert len(events) == 1
    event = events.iloc[0]
    assert event["event_type"] == "warrant_dilution"
    assert event["reported_quarter_anchor"] == "2026-Q1"
    assert event["filing_type"] == "S-3ASR / warrant exhibits"
    assert event["filing_date"] == "2026-06-22"
    assert event["downloaded_at"] == "2026-06-25T19:49:52.168707+00:00"
    assert event["accession"] == "000110465926076397"
    assert event["warrants_issued"] == 500_000.0
    assert event["potential_common_shares_issuable_max"] == 550_000.0
    assert event["exercise_price"] == pytest.approx(0.01)
    assert event["expiration_date"] == "2036-06-16"
    assert event["beneficial_ownership_limitation"] == pytest.approx(0.198)
    assert event["used_in_workbook"] == "Yes"
    assert "Valuation full-dilution sensitivity" in event["used_surfaces"]
    assert event["source_path_exists"] is True


def test_duplicate_material_and_cache_copies_collapse_to_one_event(
    tmp_path: Path,
) -> None:
    material_root = tmp_path / "tickers" / "PBI"
    cache_root = tmp_path / "sec_cache" / "PBI"
    text_by_name = {
        "d88573d8k.htm": (
            "$347 million aggregate principal amount of 6.875% Senior Notes "
            "due March 2027 were redeemed."
        ),
        "d88573dex101.htm": (
            "Incremental Term Loan A of $150 million. Total Term Loan A "
            "borrowings are $302 million and mature May 18, 2031."
        ),
        "d88573dex991.htm": (
            "The next scheduled maturity is March 2029. Existing cash and "
            "other liquidity funded fees, costs and expenses."
        ),
    }
    canonical_paths: list[Path] = []
    for name, text in text_by_name.items():
        canonical_paths.append(
            _write(material_root / "financial_statement" / name, text)
        )
        _write(cache_root / f"doc_000119312526281893_{name}", text)
    refresh_log = _write_refresh_log(
        tmp_path / "source_refresh.json",
        ticker="PBI",
        form="8-K",
        filing_date="2026-06-25",
        accession="000119312526281893",
        destinations=canonical_paths,
    )

    events = build_post_quarter_capital_events(
        ticker="PBI",
        material_roots=[material_root],
        cache_roots=[cache_root],
        source_refresh_logs=[refresh_log],
    )

    assert len(events) == 1
    assert events.iloc[0]["principal_redeemed"] == 347_000_000.0


def test_post_quarter_capital_events_writer_uses_table_filter_only(tmp_path: Path) -> None:
    wb = Workbook()
    events = pd.DataFrame(
        [
            {
                "ticker": "PBI",
                "event_key": "PBI|2026-06-23|refinancing_redemption",
                "event_type": "refinancing_redemption",
            }
        ]
    )

    write_post_quarter_capital_events_sheet(wb, events)
    workbook_path = tmp_path / "PBI_model.xlsx"
    wb.save(workbook_path)

    roundtrip = load_workbook(workbook_path)
    ws = roundtrip["PostQuarter_Capital_Events"]

    assert ws.auto_filter.ref is None
    assert "PostQuarterCapitalEvents" in ws.tables
    assert ws.tables["PostQuarterCapitalEvents"].ref == "A1:C2"
    assert [ws.cell(1, col).value for col in range(1, 4)] == [
        "ticker",
        "event_key",
        "event_type",
    ]


def test_gpre_event_is_not_emitted_without_max_issuable_share_source(
    tmp_path: Path,
) -> None:
    material_root = tmp_path / "tickers" / "GPRE"
    _write(
        material_root / "financial_statement" / "tm2618355d1_ex10-2.htm",
        "Warrants to purchase 500,000 shares of common stock.",
    )
    for suffix, amount in zip(
        ("a", "b", "c", "d"),
        (366_240, 37_120, 10_360, 86_280),
    ):
        _write(
            material_root / "financial_statement" / f"tm2618355d2_ex4-4{suffix}.htm",
            (
                f"Number of Warrants: {amount:,}. "
                "For a price of $0.01 per share, the Per Share Exercise Price. "
                "June 16, 2036 (the Expiration Date). "
                "The Beneficial Ownership Limitation shall be 19.8%."
            ),
        )

    events = build_post_quarter_capital_events(
        ticker="GPRE",
        material_roots=[material_root],
        cache_roots=[],
        source_refresh_logs=[],
    )

    assert events.empty


def test_pbi_current_debt_overlay_changes_only_a_display_copy() -> None:
    reported = pd.DataFrame(
        [
            {
                "tranche_name": "6.875% Senior Notes due March 2027",
                "amount_principal": 346_700_000.0,
                "maturity_display": "March 2027",
                "maturity_year": 2027,
                "near_term": True,
                "source_kind": "Debt_Tranches_Latest",
                "source_basis": "reported_q1",
            },
            {
                "tranche_name": "Term Loan A",
                "amount_principal": 152_000_000.0,
                "maturity_display": "May 18, 2031",
                "maturity_year": 2031,
                "near_term": False,
                "source_kind": "Debt_Tranches_Latest",
                "source_basis": "reported_q1",
            },
            {
                "tranche_name": "Senior Notes due March 2029",
                "amount_principal": 400_000_000.0,
                "maturity_display": "March 2029",
                "maturity_year": 2029,
                "near_term": False,
                "source_kind": "Debt_Tranches_Latest",
                "source_basis": "reported_q1",
            },
        ]
    )
    original = reported.copy(deep=True)
    event = {
        "ticker": "PBI",
        "event_type": "refinancing_redemption",
        "principal_redeemed": 347_000_000.0,
        "incremental_term_loan": 150_000_000.0,
        "term_loan_total": 302_000_000.0,
        "next_scheduled_maturity": "March 2029",
        "term_loan_maturity": "2031-05-18",
    }

    current = apply_pbi_current_debt_overlay(reported, event)

    pd.testing.assert_frame_equal(reported, original)
    assert not current["tranche_name"].str.contains(
        r"6\.875%.*2027",
        case=False,
        regex=True,
    ).any()
    term_loan = current[
        current["tranche_name"].str.contains("Term Loan A", case=False, na=False)
    ].iloc[0]
    assert term_loan["amount_principal"] == 302_000_000.0
    assert term_loan["maturity_display"] == "May 18, 2031"
    assert term_loan["maturity_year"] == 2031
    assert term_loan["source_kind"] == "PostQuarter_Capital_Events"
    assert term_loan["source_basis"] == "current_principal_overlay"
    assert current["tranche_name"].str.contains(
        "Senior Notes due March 2029",
        case=False,
        na=False,
    ).any()


def test_pbi_current_debt_overlay_is_noop_for_incomplete_or_wrong_event() -> None:
    reported = pd.DataFrame(
        [
            {
                "tranche_name": "6.875% Senior Notes due March 2027",
                "amount_principal": 346_700_000.0,
            }
        ]
    )

    wrong_ticker = apply_pbi_current_debt_overlay(
        reported,
        {"ticker": "GPRE", "event_type": "warrant_dilution"},
    )
    incomplete = apply_pbi_current_debt_overlay(
        reported,
        {"ticker": "PBI", "event_type": "refinancing_redemption"},
    )

    pd.testing.assert_frame_equal(wrong_ticker, reported)
    pd.testing.assert_frame_equal(incomplete, reported)


def test_pbi_current_debt_overlay_uses_bool_safe_defaults_for_sparse_tranches() -> None:
    reported = pd.DataFrame(
        [
            {
                "tranche_name": "6.875% Senior Notes due March 2027",
                "amount_principal": 346_700_000.0,
            },
            {
                "tranche_name": "Term Loan A",
                "amount_principal": 152_000_000.0,
            },
            {
                "tranche_name": "Senior Notes due March 2029",
                "amount_principal": 400_000_000.0,
            },
        ]
    )
    event = {
        "ticker": "PBI",
        "event_type": "refinancing_redemption",
        "principal_redeemed": 347_000_000.0,
        "term_loan_total": 302_000_000.0,
    }

    current = apply_pbi_current_debt_overlay(reported, event)

    unchanged = current[
        current["tranche_name"].str.contains("March 2029", case=False, na=False)
    ].iloc[0]
    assert unchanged["source_basis"] is None
    assert unchanged["source_kind"] is None
    assert unchanged["maturity_display"] is None
    assert bool(unchanged["near_term"]) is False


def test_pbi_current_debt_overlay_replaces_slide_style_reported_rows() -> None:
    reported = pd.DataFrame(
        [
            {
                "tranche_name": "Notes due March 2027",
                "amount_principal": 346_700_000.0,
                "maturity_display": "March 2027",
                "maturity_year": 2027,
            },
            {
                "tranche_name": "Term loan due March 2028",
                "amount_principal": 152_000_000.0,
                "maturity_display": "March 2028",
                "maturity_year": 2028,
            },
            {
                "tranche_name": "Notes due March 2029",
                "amount_principal": 476_000_000.0,
                "maturity_display": "March 2029",
                "maturity_year": 2029,
            },
        ]
    )
    event = {
        "ticker": "PBI",
        "event_type": "refinancing_redemption",
        "principal_redeemed": 347_000_000.0,
        "term_loan_total": 302_000_000.0,
    }

    current = apply_pbi_current_debt_overlay(reported, event)

    assert not current["tranche_name"].str.contains(
        r"notes.*2027",
        case=False,
        na=False,
        regex=True,
    ).any()
    term_rows = current[
        current["tranche_name"].str.contains("Term Loan A", case=False, na=False)
    ]
    assert len(term_rows) == 1
    assert term_rows.iloc[0]["amount_principal"] == 302_000_000.0
    assert term_rows.iloc[0]["maturity_display"] == "May 18, 2031"

