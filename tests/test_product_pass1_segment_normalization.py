from __future__ import annotations

from copy import deepcopy
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from pbi_xbrl.new_ticker_binding_planner import plan_standard_template_writes
from pbi_xbrl.new_ticker_style_planner import reproduce_style_plan
from pbi_xbrl.normalized_company_data_validation import (
    _validate_collection_business_keys,
    _validate_segment_semantics,
)
from pbi_xbrl.segment_normalization import (
    SegmentNormalizationError,
    SegmentSourceFact,
    canonical_segment_dimension_member,
    canonicalize_segment_source_facts,
)
from pbi_xbrl.standard_template_shell_identity import verify_shell_identity
from pbi_xbrl.standard_template_formula_contract import formula_target_contracts
from scripts.build_anf_shadow_normalized_package import (
    _anf_authoritative_segment_source_facts,
    _find_anf_segment_table,
    _table_member_row,
    _table_numeric_value,
    _table_scope_start,
    _table_section_rows,
    build_anf_normalized_package,
)


ROOT = Path(__file__).resolve().parents[1]


def _data_root() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = parent / "StockModelData"
        if candidate.exists():
            return candidate
    return ROOT.parent / "StockModelData"


DATA_ROOT = _data_root()
ANF_WORKBOOK = DATA_ROOT / "outputs" / "Excel stock models" / "ANF_model.xlsx"
ANF_2019_RELEASE = DATA_ROOT / "tickers" / "ANF" / "earnings_release" / "8-K_2019-03-07_earnings_release.htm"


def _source_fact(**overrides: object) -> SegmentSourceFact:
    values: dict[str, object] = {
        "metric": "revenue",
        "value": 1_452_907.0,
        "source_unit": "USD",
        "source_scale": "thousands",
        "period_type": "quarterly",
        "period": "2023-Q4",
        "dimension": "total_company",
        "member": "Total Company",
        "source_table_scope": "quarterly",
        "source_table_id": "fixture:table[2]:fourth_quarter",
        "source_row_ref": "table[2]!row[7]",
        "source_ref": "fixture.htm!table[2]!row[7]",
    }
    values.update(overrides)
    return SegmentSourceFact(**values)  # type: ignore[arg-type]


@pytest.fixture(scope="module")
def anf_package() -> dict:
    return build_anf_normalized_package(data_root=DATA_ROOT, workbook_path=ANF_WORKBOOK)


@pytest.fixture(scope="module")
def anf_plan(anf_package: dict):
    binding_payload = json.loads((ROOT / "docs" / "workbook_binding_map.json").read_text(encoding="utf-8"))
    manifest = json.loads((ROOT / "docs" / "standard_template_shell_manifest.json").read_text(encoding="utf-8"))
    identity = verify_shell_identity(
        ROOT / "templates" / "standard_stock_model_template.xlsx",
        manifest=manifest,
        binding_payload=binding_payload,
    )
    plan = plan_standard_template_writes(
        anf_package,
        binding_payload=binding_payload,
        manifest=manifest,
        shell_identity_report=identity,
    )
    assert plan.status == "PASS"
    return plan


@pytest.fixture(scope="module")
def anf_style_artifacts(anf_package: dict):
    binding_payload = json.loads((ROOT / "docs" / "workbook_binding_map.json").read_text(encoding="utf-8"))
    manifest = json.loads((ROOT / "docs" / "standard_template_shell_manifest.json").read_text(encoding="utf-8"))
    return reproduce_style_plan(
        anf_package,
        binding_payload=binding_payload,
        manifest=manifest,
        shell_path=ROOT / "templates" / "standard_stock_model_template.xlsx",
    )


def test_source_scale_and_period_scope_are_explicit() -> None:
    assert _source_fact().normalized_value == pytest.approx(1_452.907)
    assert _source_fact(
        value=3_590.1,
        source_scale="millions",
        period_type="annual",
        period="2018-FY",
        source_table_scope="annual",
    ).normalized_value == pytest.approx(3_590.1)

    with pytest.raises(SegmentNormalizationError, match="incompatible with period type"):
        _source_fact(period="2018-FY")
    with pytest.raises(SegmentNormalizationError, match="incompatible with period type"):
        _source_fact(period_type="annual", source_table_scope="annual")
    with pytest.raises(SegmentNormalizationError, match="source scope"):
        _source_fact(source_table_scope="annual")


def test_canonical_source_facts_are_order_independent_and_duplicates_fail() -> None:
    americas = _source_fact(dimension="geography", member="Americas", source_row_ref="row[4]", source_ref="fixture:row[4]")
    total = _source_fact()
    expected = canonicalize_segment_source_facts([americas, total])

    assert canonicalize_segment_source_facts([total, americas]) == expected
    duplicate = _source_fact(member="total-company", source_row_ref="row[8]", source_ref="fixture:row[8]")
    with pytest.raises(SegmentNormalizationError, match="Duplicate canonical segment business identity") as exc_info:
        canonicalize_segment_source_facts([total, duplicate])
    error = exc_info.value
    assert error.raw_pair == ("total_company", "total-company")
    assert error.canonical_pair == ("total_company", "total_company")
    assert error.source_row_ref == "row[8]"
    assert error.business_key == "quarterly|2023-Q4|total_company|total_company|revenue"
    assert "first_source_row_ref" in str(error)


@pytest.mark.parametrize(
    ("dimension", "member"),
    [
        ("brand", "Total Company"),
        ("geography", "company-total"),
        ("total_company", "Hollister"),
        ("total_company", "Americas"),
    ],
)
def test_dimension_member_pair_rejects_incompatible_total_identity(
    dimension: str,
    member: str,
) -> None:
    with pytest.raises(SegmentNormalizationError, match="requires"):
        canonical_segment_dimension_member(dimension, member)
    with pytest.raises(SegmentNormalizationError, match="requires") as exc_info:
        _source_fact(dimension=dimension, member=member)
    assert exc_info.value.raw_pair == (dimension, member)
    assert exc_info.value.canonical_pair is not None
    assert exc_info.value.source_row_ref == "table[2]!row[7]"
    assert exc_info.value.business_key


def test_total_company_aliases_cannot_evade_pair_validation_or_duplicate_detection() -> None:
    total = _source_fact()
    aliases = ("total-company", "company_total", "TOTAL")
    for index, alias in enumerate(aliases, start=8):
        duplicate = _source_fact(
            member=alias,
            source_row_ref=f"row[{index}]",
            source_ref=f"fixture:row[{index}]",
        )
        with pytest.raises(SegmentNormalizationError, match="Duplicate canonical segment business identity"):
            canonicalize_segment_source_facts([total, duplicate])

        with pytest.raises(SegmentNormalizationError, match="requires dimension 'total_company'"):
            _source_fact(
                dimension="brand" if index % 2 == 0 else "geography",
                member=alias,
                source_row_ref=f"cross-dimension-row[{index}]",
                source_ref=f"fixture:cross-dimension-row[{index}]",
            )


def test_anf_filing_table_keeps_fourth_quarter_and_full_year_scopes_separate() -> None:
    table_index, table = _find_anf_segment_table(
        ANF_2019_RELEASE,
        scope_label="Full Year",
        scale_label="(in millions)",
    )
    rows = _table_section_rows(table, "Net sales by brand", "Net sales by region")
    total_company_row = _table_member_row(table, rows, "Total company")

    assert table_index == 26
    assert _table_numeric_value(table, total_company_row, _table_scope_start(table, "Fourth Quarter")) == pytest.approx(1_155.6)
    assert _table_numeric_value(table, total_company_row, _table_scope_start(table, "Full Year")) == pytest.approx(3_590.1)


def test_anf_authoritative_segment_facts_match_independent_filing_oracles() -> None:
    facts = _anf_authoritative_segment_source_facts(DATA_ROOT)
    actual = {
        (fact.period, fact.dimension, fact.member): fact.normalized_value
        for fact in facts
    }
    assert actual == {
        ("2023-Q4", "geography", "Americas"): pytest.approx(1_191.259),
        ("2023-Q4", "geography", "EMEA"): pytest.approx(219.050),
        ("2023-Q4", "geography", "APAC"): pytest.approx(42.598),
        ("2023-Q4", "total_company", "Total Company"): pytest.approx(1_452.907),
        ("2023-Q4", "brand", "Hollister"): pytest.approx(697.704),
        ("2023-Q4", "brand", "Abercrombie"): pytest.approx(755.203),
        ("2018-FY", "total_company", "Total Company"): pytest.approx(3_590.1),
        ("2018-FY", "brand", "Hollister"): pytest.approx(2_152.5),
        ("2018-FY", "brand", "Abercrombie"): pytest.approx(1_437.6),
    }
    assert {fact.source_scale for fact in facts if fact.period == "2023-Q4"} == {"thousands"}
    assert {fact.source_scale for fact in facts if fact.period == "2018-FY"} == {"millions"}
    assert all("table[" in fact.source_ref and "row[" in fact.source_ref for fact in facts)


def test_normalized_segment_contract_is_typed_non_additive_and_duplicate_safe(anf_package: dict) -> None:
    revenue_rows = [row for row in anf_package["segments"]["items"] if row.get("metric") == "revenue"]
    assert revenue_rows
    for row in revenue_rows:
        assert row["unit"] == "$m"
        assert row["period_type"] == row["source_table_scope"]
        assert row["source_scale"] in {"ones", "thousands", "millions"}
        assert row["source_table_id"]
        assert row["source_row_ref"]
        assert row["source_ref"]

    roles = {(row["dimension"], row["aggregation_role"]) for row in revenue_rows}
    assert ("total_company", "reported_total") in roles
    assert ("geography", "dimension_member") in roles
    assert ("brand", "dimension_member") in roles

    duplicate_package = {"segments": {"items": [deepcopy(revenue_rows[0]), deepcopy(revenue_rows[0])]}}
    duplicate_package["segments"]["items"][1]["source_row_ref"] = "duplicate-row"
    issues = _validate_segment_semantics(duplicate_package)
    assert [issue.rule_id for issue in issues].count("duplicate_segment_business_identity") == 1
    duplicate_issue = next(issue for issue in issues if issue.rule_id == "duplicate_segment_business_identity")
    assert "first_raw_pair" in duplicate_issue.message
    assert "duplicate_raw_pair" in duplicate_issue.message
    assert "canonical_pair" in duplicate_issue.message
    assert "source_row_ref" in duplicate_issue.message
    assert duplicate_issue.business_row_key

    # Segment duplicates are owned by the canonical semantic validator, not the
    # generic raw collection-key pass.
    assert _validate_collection_business_keys(duplicate_package) == []


@pytest.mark.parametrize(
    ("dimension", "member", "aggregation_role"),
    [
        ("brand", "Total Company", "dimension_member"),
        ("geography", "company_total", "dimension_member"),
        ("total_company", "Hollister", "reported_total"),
    ],
)
def test_normalized_segment_pair_mutations_fail_before_planning(
    anf_package: dict,
    dimension: str,
    member: str,
    aggregation_role: str,
) -> None:
    row = deepcopy(next(item for item in anf_package["segments"]["items"] if item.get("metric") == "revenue"))
    row["dimension"] = dimension
    row["member"] = member
    row["aggregation_role"] = aggregation_role

    issues = _validate_segment_semantics({"segments": {"items": [row]}})

    assert [issue.rule_id for issue in issues] == ["invalid_segment_source_semantics"]
    assert "requires" in issues[0].message
    assert "raw_pair=" in issues[0].message
    assert "canonical_pair=" in issues[0].message
    assert "source_row_ref=" in issues[0].message
    assert "business_key=" in issues[0].message


def test_mixed_period_total_company_aliases_reproduce_exact_plan_and_style(
    anf_package: dict,
    anf_style_artifacts,
) -> None:
    canonical_value_plan, canonical_style_plan = anf_style_artifacts
    mutated = deepcopy(anf_package)
    total_revenue_rows = sorted(
        (
            row
            for row in mutated["segments"]["items"]
            if row.get("metric") == "revenue" and row.get("dimension") == "total_company"
        ),
        key=lambda row: (str(row.get("period_type") or ""), str(row.get("period") or "")),
    )
    aliases = ("Total Company", "total-company", "company_total", "TOTAL", "company total")
    assert len(total_revenue_rows) > len(aliases)
    for index, row in enumerate(total_revenue_rows):
        alias = aliases[index % len(aliases)]
        row["member"] = alias
        segment = row.get("segment")
        if isinstance(segment, dict):
            segment["value"] = alias

    binding_payload = json.loads((ROOT / "docs" / "workbook_binding_map.json").read_text(encoding="utf-8"))
    manifest = json.loads((ROOT / "docs" / "standard_template_shell_manifest.json").read_text(encoding="utf-8"))
    mutated_value_plan, mutated_style_plan = reproduce_style_plan(
        mutated,
        binding_payload=binding_payload,
        manifest=manifest,
        shell_path=ROOT / "templates" / "standard_stock_model_template.xlsx",
    )

    assert mutated_value_plan.status == "PASS"
    assert not mutated_value_plan.blocking_issues()
    assert sum(len(report.get("overflow_rows") or []) for report in mutated_value_plan.binding_reports) == 0
    assert mutated_value_plan.planned_writes == canonical_value_plan.planned_writes
    assert mutated_value_plan.binding_reports == canonical_value_plan.binding_reports
    assert mutated_value_plan.issue_ledger == canonical_value_plan.issue_ledger
    assert mutated_style_plan.actions == canonical_style_plan.actions
    assert mutated_style_plan.decisions == canonical_style_plan.decisions

    assert len(mutated_value_plan.planned_writes) == 22_214
    assert sum(len(report.get("skipped_rows") or []) for report in mutated_value_plan.binding_reports) == 2_399
    assert len(mutated_value_plan.issue_ledger["issues"]) == 761
    assert len(mutated_value_plan.issue_ledger["occurrences"]) == 2_323
    assert len(mutated_style_plan.actions) == 714
    assert len(mutated_style_plan.decisions) == 1_233


def test_exact_segment_destinations_and_missing_quarterly_disclosures(anf_plan) -> None:
    writes = {(write.target_sheet, write.target_cell): write for write in anf_plan.planned_writes}
    expected = {
        "D61": 1_191.259,
        "D62": 219.050,
        "D63": 42.598,
        "D65": 1_452.907,
        "D66": 697.704,
        "D67": 755.203,
        "B76": 3_590.1,
        "B77": 2_152.5,
        "B78": 1_437.6,
    }
    for cell, value in expected.items():
        write = writes[("BS_Segments", cell)]
        assert write.value == pytest.approx(value)
        assert "8-K_20" in write.source_ref
        assert "table[" in write.source_ref and "row[" in write.source_ref

    for cell in ("M61", "M62", "M63", "M66", "M67"):
        assert ("BS_Segments", cell) not in writes


def test_generic_segment_runtime_contains_no_anf_ticker_branch() -> None:
    source = (ROOT / "pbi_xbrl" / "segment_normalization.py").read_text(encoding="utf-8")
    assert "ANF" not in source


def test_retired_annual_financial_surface_has_no_contract_or_shell_owner() -> None:
    retired_bindings = {
        "bs_annual_financial_period_headers",
        "bs_annual_financial_revenue_series",
        "bs_annual_financial_gross_profit_series",
        "bs_annual_financial_operating_income_series",
        "bs_annual_financial_base_ebitda_series",
        "bs_annual_financial_adjusted_ebitda_series",
        "bs_annual_financial_net_income_series",
        "bs_annual_financial_operating_cash_flow_series",
        "bs_annual_financial_capital_expenditures_series",
        "bs_annual_financial_shares_outstanding_series",
        "bs_annual_financial_eps_series",
        "bs_annual_financial_total_equity_series",
        "bs_annual_financial_cash_series",
        "bs_annual_financial_debt_core_series",
    }
    retired_formula_ids = {
        "annual_gross_margin",
        "annual_operating_margin",
        "annual_ebitda_margin",
        "annual_adjusted_ebitda_margin",
        "annual_net_margin",
        "annual_free_cash_flow",
        "annual_free_cash_flow_margin",
        "annual_book_value_per_share",
        "annual_net_debt",
    }
    binding_payload = json.loads((ROOT / "docs" / "workbook_binding_map.json").read_text(encoding="utf-8"))
    module_payload = json.loads((ROOT / "docs" / "workbook_module_manifest.json").read_text(encoding="utf-8"))
    shell_manifest = json.loads((ROOT / "docs" / "standard_template_shell_manifest.json").read_text(encoding="utf-8"))
    binding_ids = {row["binding_id"] for row in binding_payload["bindings"]}
    assert retired_bindings.isdisjoint(binding_ids)
    assert retired_formula_ids.isdisjoint({row.formula_id for row in formula_target_contracts()})

    bs_blocks: list[str] = []
    bs_styles: list[str] = []
    for module in module_payload["modules"]:
        bs_blocks.extend(row["target"] for row in module.get("visible_blocks", []) if row["sheet"] == "BS_Segments")
        bs_styles.extend(row["target"] for row in module.get("style_ownership", []) if row["sheet"] == "BS_Segments")
    assert bs_blocks == ["A1:M58", "A59:M78"]
    assert bs_styles == ["A1:M58", "A59:M78"]
    assert all(range_boundaries(target)[3] <= 78 for target in bs_blocks + bs_styles)

    shell_sheet = next(row for row in shell_manifest["sheets"] if row["sheet"] == "BS_Segments")
    shell_zones = [
        row["target"]
        for zone_type in ("writable_zones", "non_writable_zones")
        for row in shell_sheet[zone_type]
    ]
    shell_contracts = [
        row["target"]
        for row in shell_manifest["planner_cell_contracts"]
        if row["sheet"] == "BS_Segments"
    ]
    assert all(range_boundaries(target)[3] <= 78 for target in shell_zones + shell_contracts)

    workbook = load_workbook(ROOT / "templates" / "standard_stock_model_template.xlsx", data_only=False)
    try:
        sheet = workbook["BS_Segments"]
        assert retired_bindings.isdisjoint(set(workbook.defined_names))
        for row in range(81, 105):
            assert sheet.row_dimensions[row].hidden is True
            for column in range(1, 14):
                cell = sheet.cell(row, column)
                assert cell.value is None
                assert cell.protection.locked is True
    finally:
        workbook.close()
