from __future__ import annotations

from copy import deepcopy
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from pbi_xbrl.new_ticker_guidance_scope import (
    GuidanceProjectionError,
    build_valuation_guidance_projection,
)
from pbi_xbrl.new_ticker_thesis_projection import (
    ThesisProjectionError,
    build_valuation_thesis_projection,
)


ROOT = Path(__file__).resolve().parents[1]
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
MODULE_MANIFEST = ROOT / "docs" / "workbook_module_manifest.json"
SHELL_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
SHELL_WORKBOOK = ROOT / "templates" / "standard_stock_model_template.xlsx"


def _field(value: object, *, unit: str = "", source_ref: str) -> dict:
    row = {
        "value": value,
        "status": "populated",
        "source_ref": source_ref,
        "core": True,
    }
    if unit:
        row["unit"] = unit
    return row


def _guidance_row(
    metric: str,
    horizon: str,
    unit: str,
    role: str,
    priority: int,
    evidence_key: str,
    *,
    publication_date: str,
    value: str | None = None,
    review_state: str = "accepted",
) -> dict:
    source_ref = f"fixture:{evidence_key}"
    return {
        "metric": _field(metric, source_ref=source_ref),
        "value": _field(value or f"guide:{evidence_key}", unit=unit, source_ref=source_ref),
        "horizon": _field(horizon, source_ref=source_ref),
        "publication_date": publication_date,
        "stated_in_period": "2025-Q4",
        "source_date": "2026-01-31",
        "update_stage": "initial",
        "display_role": role,
        "display_priority": priority,
        "visibility_disposition": "historical" if role == "history" else "visible",
        "review_state": review_state,
        "progress_status": "resolved_pass" if role == "history" else "open",
        "evidence_key": evidence_key,
        "evidence_refs": [source_ref],
        "source_ref": source_ref,
    }


def _guidance_fixture() -> list[dict]:
    rows = [
        _guidance_row("Revenue", "2026 year", "%", "current_primary", 1, "p-revenue-fy", publication_date="2026-03-04"),
        _guidance_row("Revenue", "2026-Q1", "%", "current_primary", 2, "p-revenue-q", publication_date="2026-03-04"),
        _guidance_row("Operating margin", "2026 year", "%", "current_primary", 3, "p-margin-fy", publication_date="2026-03-04"),
        _guidance_row("Operating margin", "2026-Q1", "%", "current_primary", 4, "p-margin-q", publication_date="2026-03-04"),
        _guidance_row("Adj EPS", "2026 year", "$/share", "current_primary", 5, "p-eps-fy", publication_date="2026-03-04"),
        _guidance_row("Adj EPS", "2026-Q1", "$/share", "current_primary", 6, "p-eps-q", publication_date="2026-03-04"),
        _guidance_row("Real estate activity", "2026 year", "stores", "current_primary", 7, "p-real-estate-fy", publication_date="2026-03-04"),
        _guidance_row("Capex", "2026 year", "$m", "current_secondary", 102, "s-capex-fy", publication_date="2026-03-04"),
        _guidance_row("Diluted shares", "2026-Q1", "m shares", "current_secondary", 103, "s-shares-q", publication_date="2026-03-04"),
        _guidance_row("Diluted shares", "2026 year", "million_shares", "current_secondary", 104, "s-shares-fy", publication_date="2026-03-04"),
        _guidance_row("Real-estate activity", "2026-Q1", "stores", "current_secondary", 107, "s-real-estate-q", publication_date="2026-03-04"),
        _guidance_row("Share repurchases", "2026-Q1", "$m", "current_secondary", 111, "s-buyback-q", publication_date="2026-03-04"),
        _guidance_row("Share repurchases", "2026 year", "$m", "current_secondary", 112, "s-buyback-fy", publication_date="2026-03-04"),
    ]
    history = [
        ("Revenue", "%", "h-revenue"),
        ("Operating margin", "%", "h-margin"),
        ("Adjusted EPS", "$/share", "h-eps"),
        ("Capital expenditures", "$m", "h-capex"),
        ("Diluted shares", "shares_m", "h-shares"),
        ("Real estate activity", "stores", "h-real-estate"),
        ("Share repurchases", "$m", "h-buyback"),
    ]
    rows.extend(
        _guidance_row(metric, "FY2025", unit, "history", 999, evidence, publication_date="2026-01-12")
        for metric, unit, evidence in history
    )
    older = _guidance_row("Revenue", "FY2024", "%", "history", 999, "h-revenue-old", publication_date="2025-01-10")
    rows.append(older)
    audit_only = _guidance_row("Revenue", "FY2025", "%", "audit_only", 999, "audit-only", publication_date="2026-01-13")
    rows.append(audit_only)
    return rows


def test_guidance_projection_is_exact_and_row_order_independent() -> None:
    rows = _guidance_fixture()

    first = build_valuation_guidance_projection(rows, profile_pack_ids={"retail_operating_pack"})
    second = build_valuation_guidance_projection(list(reversed(rows)), profile_pack_ids={"retail_operating_pack"})

    assert first.to_dict() == second.to_dict()
    assert [row.canonical_metric for row in first.current_primary_rows] == [
        "revenue",
        "revenue",
        "operating_margin",
        "operating_margin",
        "adjusted_eps",
        "adjusted_eps",
        "real_estate_activity",
    ]
    assert [(row.canonical_metric, row.horizon) for row in first.current_secondary_rows] == [
        ("capital_expenditures", "FY2026"),
        ("diluted_shares", "2026-Q1"),
        ("diluted_shares", "FY2026"),
        ("real_estate_activity", "2026-Q1"),
        ("share_repurchases", "2026-Q1"),
        ("share_repurchases", "FY2026"),
    ]
    assert [row.canonical_metric for row in first.historical_rows] == [
        "revenue",
        "operating_margin",
        "adjusted_eps",
        "capital_expenditures",
        "diluted_shares",
        "real_estate_activity",
        "share_repurchases",
    ]
    assert first.current_secondary_rows[1].unit == "shares_m"
    assert first.historical_rows[4].unit == "shares_m"
    assert {row.display_state for row in first.historical_rows} == {"history / accepted"}
    assert "resolved_pass" not in json.dumps(first.to_dict(), sort_keys=True)


def test_guidance_projection_conflicts_and_invalid_visible_aliases_fail_closed() -> None:
    rows = _guidance_fixture()
    conflict = deepcopy(rows[0])
    conflict["evidence_key"] = "conflicting-revenue"
    conflict["evidence_refs"] = ["fixture:conflicting-revenue"]
    conflict["source_ref"] = "fixture:conflicting-revenue"
    conflict["display_priority"] = 8
    conflict["value"] = _field("conflicting guide", unit="%", source_ref="fixture:conflicting-revenue")

    with pytest.raises(GuidanceProjectionError, match="guidance_projection_conflict"):
        build_valuation_guidance_projection([*rows, conflict], profile_pack_ids={"retail_operating_pack"})

    for mutation, expected in (
        (("metric", "Revenue-ish"), "unknown_visible_guidance_metric"),
        (("unit", "dollars-ish"), "unknown_visible_guidance_unit"),
        (("horizon", "soon"), "unknown_visible_guidance_horizon"),
        (("evidence", None), "guidance_evidence_missing"),
    ):
        invalid = deepcopy(rows)
        if mutation[0] == "metric":
            invalid[0]["metric"]["value"] = mutation[1]
        elif mutation[0] == "unit":
            invalid[0]["value"]["unit"] = mutation[1]
        elif mutation[0] == "horizon":
            invalid[0]["horizon"]["value"] = mutation[1]
        else:
            invalid[0]["evidence_refs"] = []
        with pytest.raises(GuidanceProjectionError, match=expected):
            build_valuation_guidance_projection(invalid, profile_pack_ids={"retail_operating_pack"})


def test_guidance_projection_priority_supersession_status_and_profile_rules() -> None:
    rows = _guidance_fixture()

    duplicate_priority = deepcopy(rows)
    duplicate_priority[2]["display_priority"] = duplicate_priority[0]["display_priority"]
    with pytest.raises(GuidanceProjectionError, match="duplicate_guidance_display_priority"):
        build_valuation_guidance_projection(
            duplicate_priority,
            profile_pack_ids={"retail_operating_pack"},
        )

    rejected_status = deepcopy(rows)
    rejected_status[0]["review_state"] = "manual_review_required"
    with pytest.raises(GuidanceProjectionError, match="guidance_source_status_rejected"):
        build_valuation_guidance_projection(
            rejected_status,
            profile_pack_ids={"retail_operating_pack"},
        )

    replacement = deepcopy(rows[0])
    replacement_ref = "fixture:p-revenue-fy-replacement"
    replacement["evidence_key"] = "p-revenue-fy-replacement"
    replacement["evidence_refs"] = [replacement_ref]
    replacement["source_ref"] = replacement_ref
    replacement["metric"]["source_ref"] = replacement_ref
    replacement["value"] = _field("replacement guide", unit="%", source_ref=replacement_ref)
    replacement["horizon"]["source_ref"] = replacement_ref
    replacement["publication_date"] = "2026-03-05"
    replacement["display_priority"] = 8
    replacement["supersedes_evidence_keys"] = [rows[0]["evidence_key"]]
    superseded = build_valuation_guidance_projection(
        [*rows, replacement],
        profile_pack_ids={"retail_operating_pack"},
    )
    assert superseded.current_primary_rows[0].evidence_key == "p-revenue-fy-replacement"
    assert any(
        row["evidence_key"] == rows[0]["evidence_key"]
        and row["disposition"] == "superseded_or_withdrawn"
        for row in superseded.selection_audit
    )

    non_retail = build_valuation_guidance_projection(
        rows,
        profile_pack_ids={"commodity_ethanol_pack"},
    )
    assert all(
        row.canonical_metric != "real_estate_activity"
        for row in (
            *non_retail.current_primary_rows,
            *non_retail.current_secondary_rows,
            *non_retail.historical_rows,
        )
    )
    assert sum(
        row["disposition"] == "profile_slot_inactive"
        for row in non_retail.selection_audit
    ) == 3


def _thesis_field(value: str, review_state: str, key: str) -> dict:
    return {
        "value": value,
        "status": "populated",
        "source_ref": f"fixture:{key}",
        "evidence_refs": [f"fixture:{key}"],
        "evidence_classification": "source_backed_fact" if review_state == "accepted" else "analyst_interpretation_requiring_review",
        "review_state": review_state,
    }


def _investment_case_fixture() -> dict:
    return {
        "key_debate": _thesis_field("Key debate", "manual_review_required", "key-debate"),
        "why_it_can_work": _thesis_field("Why it can work", "accepted", "why"),
        "upside_factors": _thesis_field("Upside", "manual_review_required", "upside"),
        "downside_factors": _thesis_field("Downside", "manual_review_required", "downside"),
        "watch_next": _thesis_field("Watch next", "manual_review_required", "watch"),
        "current_stance": _thesis_field("Current stance", "manual_review_required", "stance"),
        "invalidators": [
            {
                "business_key": "sales-execution-breaks",
                "text": _thesis_field("Sales invalidator", "manual_review_required", "sales"),
                "display_order": 1,
            },
            {
                "business_key": "margin-durability-breaks",
                "text": _thesis_field("Margin invalidator", "manual_review_required", "margin"),
                "display_order": 2,
            },
        ],
    }


def test_thesis_projection_uses_exact_typed_fields_and_lineage() -> None:
    projection = build_valuation_thesis_projection(_investment_case_fixture())

    assert [row.item_id for row in projection.rows] == [
        "key_debate",
        "why_it_can_work",
        "upside_factors",
        "downside_factors",
        "watch_next",
        "current_stance",
        "sales-execution-breaks",
        "margin-durability-breaks",
    ]
    assert [row.review_state for row in projection.rows] == [
        "manual_review_required",
        "accepted",
        "manual_review_required",
        "manual_review_required",
        "manual_review_required",
        "manual_review_required",
        "manual_review_required",
        "manual_review_required",
    ]
    assert projection.rows[1].normalized_path == "investment_case.why_it_can_work"
    assert projection.rows[6].normalized_path == "investment_case.invalidators.sales-execution-breaks"
    assert all(row.source_ref and row.evidence_refs for row in projection.rows)

    malformed = _investment_case_fixture()
    malformed["key_debate"] = "infer this prose"
    with pytest.raises(ThesisProjectionError, match="must be a typed normalized field"):
        build_valuation_thesis_projection(malformed)


def test_product_pass2b_binding_and_ownership_contracts_are_bounded() -> None:
    binding_payload = json.loads(BINDING_MAP.read_text(encoding="utf-8"))
    bindings = {row["binding_id"]: row for row in binding_payload["bindings"]}
    assert {
        "valuation_guidance_current_primary_rows",
        "valuation_guidance_current_secondary_rows",
        "valuation_guidance_historical_rows",
        "valuation_thesis_debate_rows",
    } <= set(bindings)
    assert not {
        "valuation_guidance_rows",
        "valuation_guidance_rows_lower",
        "valuation_guidance_status_rows",
        "valuation_guidance_status_rows_lower",
        "valuation_operating_driver_sidecar_rows",
        "valuation_thesis_bridge_rows",
    } & set(bindings)
    assert all(
        ".items.0" not in json.dumps(bindings[binding_id], sort_keys=True)
        for binding_id in (
            "valuation_guidance_current_primary_rows",
            "valuation_guidance_current_secondary_rows",
            "valuation_guidance_historical_rows",
        )
    )

    modules = json.loads(MODULE_MANIFEST.read_text(encoding="utf-8"))["modules"]
    by_module = {row["module_id"]: row for row in modules}
    assert "valuation_operating_driver_sidecar" not in {
        row["block_id"] for row in by_module["operating_drivers"]["visible_blocks"]
    }
    assert "valuation_thesis_debate" in {
        row["block_id"] for row in by_module["investment_case_market_implied"]["visible_blocks"]
    }

    shell = json.loads(SHELL_MANIFEST.read_text(encoding="utf-8"))
    valuation = next(row for row in shell["sheets"] if row["sheet"] == "Valuation")
    owned_ranges = [zone["target"] for zone in [*valuation["writable_zones"], *valuation["non_writable_zones"]]]
    assert "O37:AC47" not in owned_ranges
    assert not any(target in owned_ranges for target in ("A145:M145", "A151:M151", "A158:M188"))


def _assert_blank_locked_range(worksheet, target: str) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(target)
    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            assert cell.value is None, cell.coordinate
            assert cell.protection.locked, cell.coordinate


def test_product_pass2b_checked_in_shell_retires_inactive_capacity() -> None:
    workbook = load_workbook(SHELL_WORKBOOK, data_only=False)
    try:
        worksheet = workbook["Valuation"]

        _assert_blank_locked_range(worksheet, "O22:AC25")
        _assert_blank_locked_range(worksheet, "O37:AC47")
        _assert_blank_locked_range(worksheet, "A145:M151")
        _assert_blank_locked_range(worksheet, "A169:M188")
        assert all(worksheet.row_dimensions[row_idx].hidden for row_idx in range(145, 152))
        assert all(not worksheet.row_dimensions[row_idx].hidden for row_idx in range(152, 169))
        assert all(worksheet.row_dimensions[row_idx].hidden for row_idx in range(169, 189))

        retired_ranges = ("O22:AC25", "O37:AC47", "A145:M151", "A169:M188")
        for merged_range in worksheet.merged_cells.ranges:
            merged = str(merged_range)
            assert not any(
                _ranges_intersect_for_test(merged, retired)
                for retired in retired_ranges
            ), merged
        for validation in worksheet.data_validations.dataValidation:
            for cell_range in validation.ranges.ranges:
                assert not any(
                    _ranges_intersect_for_test(str(cell_range), retired)
                    for retired in retired_ranges
                ), cell_range

        assert sum(
            1
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ) == 2_609
        assert len(workbook.worksheets) == 44
        assert all(sheet.protection.sheet for sheet in workbook.worksheets)
    finally:
        workbook.close()


def _ranges_intersect_for_test(left: str, right: str) -> bool:
    left_min_col, left_min_row, left_max_col, left_max_row = range_boundaries(left)
    right_min_col, right_min_row, right_max_col, right_max_row = range_boundaries(right)
    return not (
        left_max_col < right_min_col
        or right_max_col < left_min_col
        or left_max_row < right_min_row
        or right_max_row < left_min_row
    )
