from __future__ import annotations

import ast
from copy import deepcopy
import json
from pathlib import Path

from openpyxl import load_workbook
import pytest

from pbi_xbrl.new_ticker_debt_projection import (
    DebtFacilityProjectionPolicy,
    DebtFacilityProjectionRole,
    DebtProjectionError,
    build_debt_workbook_projection,
)
from pbi_xbrl.new_ticker_style_planner import reproduce_style_plan
from pbi_xbrl.new_ticker_value_filler import fill_standard_template_from_package


ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = ROOT.parents[2] / "StockModelData"
ANF_PACKAGE = DATA_ROOT / "outputs" / "stress_tests" / "ANF_new_ticker_engine" / "ANF_normalized_data_package.json"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
SHELL_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
PROJECTION_MODULE = ROOT / "pbi_xbrl" / "new_ticker_debt_projection.py"


def _json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _amount(value: float, *, as_of: str = "2026-05-02") -> dict[str, object]:
    return {
        "value": value,
        "status": "populated",
        "currency": "USD",
        "unit": "$m",
        "source_value": value * 1_000,
        "source_unit": "USD",
        "source_scale": "thousands",
        "as_of_date": as_of,
        "source_ref": f"fixture.htm#amount-{value}",
        "source_row_ref": f"fixture-table:{value}",
        "evidence_refs": [f"fixture.htm#amount-{value}"],
        "evidence_classification": "source_backed_fact",
        "derivation": "",
        "reason": "",
        "core": False,
    }


def _instrument(index: int, *, balance: float = 100.0) -> dict[str, object]:
    instrument_id = f"term_loan_{index}"
    return {
        "instrument_id": instrument_id,
        "instrument_name": f"Term loan {index}",
        "instrument_type": "term_loan",
        "issuer": "Fixture issuer",
        "currency": "USD",
        "as_of_date": "2026-05-02",
        "publication_date": "2026-06-05",
        "period_role": "current",
        "source_status": "accepted",
        "source_table_scope": "debt_note_table",
        "aggregation_role": "core_debt",
        "balance": _amount(balance),
        "current_balance": _amount(balance * 0.25),
        "noncurrent_balance": _amount(balance * 0.75),
        "rate_type": "fixed",
        "reference_rate": "",
        "spread_bps": None,
        "effective_rate": 6.25,
        "maturity_date": f"203{index}-05-02",
        "secured_status": "secured",
        "seniority": "senior",
        "evidence_key": f"fixture_{instrument_id}",
        "evidence_refs": [f"fixture.htm#{instrument_id}"],
        "source_refs": [f"fixture.htm#{instrument_id}"],
        "source_row_ref": f"debt-table:row-{index}",
        "source_document_sha256": str(index) * 64,
        "reason": "Source-backed fixture.",
    }


def _maturity(index: int, *, amount: float = 100.0) -> dict[str, object]:
    instrument_id = f"term_loan_{index}"
    return {
        "maturity_id": f"{instrument_id}_203{index}",
        "instrument_id": instrument_id,
        "maturity_type": "contractual_principal",
        "due_date": f"203{index}-05-02",
        "maturity_bucket": f"fy203{index}",
        "currency": "USD",
        "as_of_date": "2026-05-02",
        "publication_date": "2026-06-05",
        "period_role": "current",
        "source_status": "accepted",
        "source_table_scope": "debt_maturity_table",
        "aggregation_role": "core_debt_maturity",
        "amount": _amount(amount),
        "evidence_key": f"fixture_{instrument_id}_maturity",
        "evidence_refs": [f"fixture.htm#{instrument_id}-maturity"],
        "source_refs": [f"fixture.htm#{instrument_id}-maturity"],
        "source_row_ref": f"maturity-table:row-{index}",
        "source_document_sha256": str(index + 4) * 64,
        "reason": "Source-backed fixture.",
    }


def _funded_debt_package(*, mismatch: bool = False) -> dict[str, object]:
    instruments = [_instrument(index) for index in range(1, 5)]
    maturities = [_maturity(index, amount=99.0 if mismatch and index == 4 else 100.0) for index in range(1, 5)]
    return {
        "debt_liquidity": {
            "facilities": [],
            "instruments": instruments,
            "maturities": maturities,
            "credit_notes": [],
        },
        "quarterly_financials": {"rows": []},
    }


def _set_maturity_snapshot(row: dict[str, object], as_of_date: str, *, period_role: str = "current") -> None:
    row["as_of_date"] = as_of_date
    row["period_role"] = period_role
    amount = row["amount"]
    assert isinstance(amount, dict)
    amount["as_of_date"] = as_of_date


def _with_second_current_facility(package: dict, *, facility_id: str = "Secondary Revolver") -> dict:
    mutated = deepcopy(package)
    facilities = mutated["debt_liquidity"]["facilities"]
    latest = max(facilities, key=lambda row: (row["as_of_date"], row["facility_id"]))
    duplicate = deepcopy(latest)
    duplicate["facility_id"] = facility_id
    duplicate["facility_name"] = "Secondary Revolver"
    duplicate["evidence_key"] = "fixture_secondary_revolver"
    duplicate["evidence_refs"] = ["fixture.htm#secondary-revolver"]
    duplicate["source_refs"] = ["fixture.htm#secondary-revolver"]
    duplicate["source_row_ref"] = "fixture-facility-table:secondary-revolver"
    duplicate["source_document_sha256"] = "e" * 64
    for field in (
        "commitment",
        "loan_cap",
        "drawn_balance",
        "letters_of_credit",
        "gross_capacity",
        "minimum_excess_availability",
        "net_availability",
        "cash_and_equivalents",
        "restricted_cash",
        "same_date_liquidity",
    ):
        amount = duplicate[field]
        amount["source_ref"] = f"fixture.htm#secondary-revolver-{field}"
        amount["source_row_ref"] = f"fixture-facility-table:secondary-revolver:{field}"
        amount["evidence_refs"] = [f"fixture.htm#secondary-revolver-{field}"]
    facilities.append(duplicate)
    return mutated


def _facility_policy(
    *,
    primary_liquidity_role: str,
    secondary_liquidity_role: str,
) -> DebtFacilityProjectionPolicy:
    return DebtFacilityProjectionPolicy(
        roles=(
            DebtFacilityProjectionRole(
                facility_id="anf_abl_facility",
                profile_role="primary",
                liquidity_role=primary_liquidity_role,
                aggregation_group_id="fixture_corporate_liquidity",
                evidence_refs=("fixture-policy.htm#anf-primary",),
            ),
            DebtFacilityProjectionRole(
                facility_id="secondary_revolver",
                profile_role="detail",
                liquidity_role=secondary_liquidity_role,
                aggregation_group_id="fixture_corporate_liquidity",
                evidence_refs=("fixture-policy.htm#secondary-role",),
            ),
        )
    )


@pytest.fixture(scope="module")
def anf_package() -> dict:
    if not ANF_PACKAGE.exists():
        pytest.skip("ANF normalized debt fixture is unavailable.")
    return _json(ANF_PACKAGE)


def test_anf_debt_product_rows_are_exact_and_row_order_independent(anf_package: dict) -> None:
    projection = build_debt_workbook_projection(anf_package)
    reordered = deepcopy(anf_package)
    for collection in ("facilities", "instruments", "credit_notes"):
        reordered["debt_liquidity"][collection].reverse()

    assert projection.to_dict() == build_debt_workbook_projection(reordered).to_dict()
    assert projection.projection_digest == "d22833a1043e3969db221eeef5de5bbf54b94dde954ac1602a9da0fe4b4fda35"
    assert len(projection.debt_profile_rows) == 11
    assert [row.value for row in projection.debt_profile_rows[:10]] == pytest.approx(
        [500.0, 500.0, 0.0, 0.469, 499.531, 50.0, 449.531, 594.08, 7.336, 1292.477]
    )
    assert projection.debt_profile_rows[2].state == "reported_zero"
    assert projection.debt_profile_rows[8].expiry_or_maturity == ""
    assert projection.debt_profile_rows[9].expiry_or_maturity == ""
    assert projection.debt_profile_rows[10].value is None
    assert projection.debt_profile_rows[10].state == "unavailable"
    assert projection.debt_profile_rows[10].definition_or_source == "no source-backed core borrowings identified"
    assert projection.debt_profile_rows[10].as_of_date == "2026-05-02"

    assert len(projection.revolver_history_rows) == 12
    assert [row.as_of_date for row in projection.revolver_history_rows] == sorted(
        row.as_of_date for row in projection.revolver_history_rows
    )
    missing = [row for row in projection.revolver_history_rows if row.drawn is None]
    assert [row.as_of_date for row in missing] == ["2024-08-03", "2024-11-02"]
    assert all(row.source_state == "accepted_with_missing_drawn" for row in missing)
    assert all(row.drawn == 0.0 for row in projection.revolver_history_rows if row not in missing)

    by_period = {row.period: row for row in projection.leverage_liquidity_rows}
    assert by_period["2026-Q1"].cash == pytest.approx(594.08)
    assert by_period["2026-Q1"].revolver_availability == pytest.approx(449.531)
    assert by_period["2025-Q4"].cash == pytest.approx(759.54)
    assert by_period["2025-Q4"].revolver_availability == pytest.approx(449.546)
    assert all(row.core_debt is None and row.disposition_state == "debt_unavailable" for row in by_period.values())

    assert [(row.priority, row.topic) for row in projection.debt_credit_note_rows] == [
        (5, "Covenant compliance"),
        (6, "Senior-notes redemption / no ABL borrowings"),
    ]
    assert projection.debt_maturity_rows == ()
    assert dict(projection.sheet_states) == {
        "Debt_Profile": "visible",
        "Revolver_History": "visible",
        "Leverage_Liquidity": "visible",
        "Debt_Credit_Notes": "visible",
        "Debt_Maturity_Ladder": "hidden",
        "Debt_Tranches_Latest": "hidden",
        "Debt_Tranches_Q": "hidden",
        "Debt_Buckets": "hidden",
        "Debt_Recon": "hidden",
    }


def test_funded_debt_maturity_visibility_requires_exact_reconciliation() -> None:
    accepted = build_debt_workbook_projection(_funded_debt_package())
    assert len(accepted.debt_profile_rows) == 5
    assert len(accepted.debt_maturity_rows) == 4
    assert dict(accepted.sheet_states)["Debt_Profile"] == "visible"
    assert dict(accepted.sheet_states)["Debt_Maturity_Ladder"] == "visible"

    failed = build_debt_workbook_projection(_funded_debt_package(mismatch=True))
    assert failed.debt_maturity_rows == ()
    assert dict(failed.sheet_states)["Debt_Maturity_Ladder"] == "hidden"
    assert [row["rule_id"] for row in failed.blocking_issues] == ["debt_maturity_reconciliation_failed"]


def test_maturity_snapshot_must_match_its_instrument_even_when_amounts_reconcile() -> None:
    package = _funded_debt_package()
    for row in package["debt_liquidity"]["maturities"]:
        _set_maturity_snapshot(row, "2026-04-30")

    projection = build_debt_workbook_projection(package)

    assert projection.debt_maturity_rows == ()
    assert dict(projection.sheet_states)["Debt_Maturity_Ladder"] == "hidden"
    assert {row["rule_id"] for row in projection.blocking_issues} == {
        "debt_maturity_instrument_period_mismatch"
    }
    assert all(row["instrument_as_of_date"] == "2026-05-02" for row in projection.blocking_issues)
    assert all(row["maturity_as_of_dates"] == ["2026-04-30"] for row in projection.blocking_issues)


def test_mixed_maturity_snapshots_block_the_complete_schedule() -> None:
    package = _funded_debt_package()
    current = package["debt_liquidity"]["maturities"][0]
    current["amount"] = _amount(50.0)
    prior = deepcopy(current)
    prior["maturity_id"] = "term_loan_1_prior_snapshot"
    prior["evidence_key"] = "fixture_term_loan_1_prior_maturity"
    prior["evidence_refs"] = ["fixture.htm#term-loan-1-prior-maturity"]
    prior["source_refs"] = ["fixture.htm#term-loan-1-prior-maturity"]
    prior["source_row_ref"] = "maturity-table:prior-row-1"
    prior["source_document_sha256"] = "f" * 64
    prior["amount"] = _amount(50.0, as_of="2026-04-30")
    _set_maturity_snapshot(prior, "2026-04-30", period_role="historical")
    package["debt_liquidity"]["maturities"].append(prior)

    projection = build_debt_workbook_projection(package)

    assert projection.debt_maturity_rows == ()
    assert dict(projection.sheet_states)["Debt_Maturity_Ladder"] == "hidden"
    assert [row["rule_id"] for row in projection.blocking_issues] == [
        "debt_maturity_instrument_period_mismatch"
    ]
    assert projection.blocking_issues[0]["maturity_business_keys"] == [
        "maturity|term_loan_1_prior_snapshot|2026-04-30"
    ]


def test_missing_maturity_snapshot_blocks_before_projection_rowsets() -> None:
    package = _funded_debt_package()
    maturity = package["debt_liquidity"]["maturities"][0]
    _set_maturity_snapshot(maturity, "")

    with pytest.raises(DebtProjectionError) as exc_info:
        build_debt_workbook_projection(package)

    assert exc_info.value.rule_id == "invalid_debt_date"
    assert exc_info.value.context["field"] == "as_of_date"


def test_maturity_publication_date_does_not_replace_snapshot_identity() -> None:
    package = _funded_debt_package()
    for index, row in enumerate(package["debt_liquidity"]["maturities"], start=10):
        row["publication_date"] = f"2026-06-{index:02d}"

    projection = build_debt_workbook_projection(package)

    assert not projection.blocking_issues
    assert len(projection.debt_maturity_rows) == 4
    assert dict(projection.sheet_states)["Debt_Maturity_Ladder"] == "visible"


def test_ambiguous_same_date_facilities_block_profile_and_period_liquidity(anf_package: dict) -> None:
    package = _with_second_current_facility(anf_package)
    projection = build_debt_workbook_projection(package)
    reordered = deepcopy(package)
    reordered["debt_liquidity"]["facilities"].reverse()

    assert projection.to_dict() == build_debt_workbook_projection(reordered).to_dict()
    assert projection.debt_profile_rows == ()
    assert dict(projection.sheet_states)["Debt_Profile"] == "hidden"
    assert dict(projection.sheet_states)["Leverage_Liquidity"] == "hidden"
    assert [row["rule_id"] for row in projection.blocking_issues] == [
        "debt_profile_ambiguous_current_facility",
        "debt_liquidity_ambiguous_facility_aggregation",
    ]
    assert not any(row.as_of_date == "2026-05-02" for row in projection.leverage_liquidity_rows)
    assert len({row.as_of_date for row in projection.leverage_liquidity_rows}) == len(
        projection.leverage_liquidity_rows
    )
    assert len({row.period for row in projection.leverage_liquidity_rows}) == len(
        projection.leverage_liquidity_rows
    )


def test_explicit_primary_facility_is_stable_across_order_and_id_spelling(anf_package: dict) -> None:
    package = _with_second_current_facility(anf_package)
    policy = _facility_policy(
        primary_liquidity_role="primary",
        secondary_liquidity_role="overlapping_detail",
    )
    projection = build_debt_workbook_projection(package, facility_policy=policy)

    reordered = deepcopy(package)
    primary = next(row for row in reordered["debt_liquidity"]["facilities"] if row["facility_id"] == "anf_abl_facility")
    primary["facility_id"] = "ANF ABL Facility"
    reordered["debt_liquidity"]["facilities"].reverse()
    reordered_projection = build_debt_workbook_projection(reordered, facility_policy=policy)

    assert not projection.blocking_issues
    assert projection.to_dict() == reordered_projection.to_dict()
    assert {row.facility_or_instrument for row in projection.debt_profile_rows[:7]} == {"ABL Facility"}
    may = [row for row in projection.leverage_liquidity_rows if row.as_of_date == "2026-05-02"]
    assert len(may) == 1
    assert may[0].revolver_availability == pytest.approx(449.531)
    assert len(projection.revolver_history_rows) == 12
    assert all(row.facility == "ABL Facility" for row in projection.revolver_history_rows)


def test_explicit_additive_facilities_create_one_evidence_backed_period_row(anf_package: dict) -> None:
    package = _with_second_current_facility(anf_package)
    policy = _facility_policy(
        primary_liquidity_role="additive",
        secondary_liquidity_role="additive",
    )

    projection = build_debt_workbook_projection(package, facility_policy=policy)
    reordered = deepcopy(package)
    reordered["debt_liquidity"]["facilities"].reverse()
    reordered_projection = build_debt_workbook_projection(reordered, facility_policy=policy)
    may = [row for row in projection.leverage_liquidity_rows if row.as_of_date == "2026-05-02"]

    assert not projection.blocking_issues
    assert projection.to_dict() == reordered_projection.to_dict()
    assert len(may) == 1
    assert may[0].cash == pytest.approx(594.08)
    assert may[0].revolver_availability == pytest.approx(899.062)
    assert "anf_abl_facility" in may[0].component_period_explanation
    assert "secondary_revolver" in may[0].component_period_explanation
    assert "fixture_secondary_revolver" in may[0].evidence_key
    assert "fixture-policy.htm#secondary-role" in may[0].evidence_key
    assert len({row.period for row in projection.leverage_liquidity_rows}) == 12


def test_incompatible_declared_facility_aggregation_fails_closed(anf_package: dict) -> None:
    package = _with_second_current_facility(anf_package)
    policy = _facility_policy(
        primary_liquidity_role="primary",
        secondary_liquidity_role="additive",
    )

    projection = build_debt_workbook_projection(package, facility_policy=policy)

    assert [row["rule_id"] for row in projection.blocking_issues] == [
        "debt_liquidity_ambiguous_facility_aggregation"
    ]
    assert not any(row.as_of_date == "2026-05-02" for row in projection.leverage_liquidity_rows)
    assert len({row.as_of_date for row in projection.leverage_liquidity_rows}) == len(
        projection.leverage_liquidity_rows
    )


def test_no_data_and_disabled_debt_module_leave_every_analytical_sheet_hidden() -> None:
    package = {"debt_liquidity": {"facilities": [], "instruments": [], "maturities": [], "credit_notes": []}}
    for projection in (
        build_debt_workbook_projection(package),
        build_debt_workbook_projection(package, debt_module_active=False),
    ):
        assert not any(state == "visible" for _sheet, state in projection.sheet_states)
        assert not projection.debt_profile_rows
        assert not projection.debt_maturity_rows


def test_projection_has_no_ticker_branch_items_zero_or_prose_scoring() -> None:
    source = PROJECTION_MODULE.read_text(encoding="utf-8")
    assert "items.0" not in source
    assert "sentiment" not in source.casefold()
    tree = ast.parse(source)
    assert all("ticker" not in ast.unparse(node.test).casefold() for node in ast.walk(tree) if isinstance(node, ast.If))


def test_anf_binding_style_and_visibility_projection_is_exact(anf_package: dict) -> None:
    value_plan, style_plan = reproduce_style_plan(
        anf_package,
        binding_payload=_json(BINDING_MAP),
        manifest=_json(SHELL_MANIFEST),
        shell_path=SHELL,
    )
    assert value_plan.status == "PASS"
    assert {
        binding_id: sum(write.binding_id == binding_id for write in value_plan.planned_writes)
        for binding_id in (
            "debt_profile_resolved_rows",
            "revolver_history_resolved_rows",
            "revolver_history_companion_rows",
            "leverage_liquidity_resolved_rows",
            "leverage_liquidity_availability_rows",
            "leverage_liquidity_companion_rows",
            "debt_credit_notes_resolved_rows",
            "debt_maturity_ladder_resolved_rows",
        )
    } == {
        "debt_profile_resolved_rows": 103,
        "revolver_history_resolved_rows": 118,
        "revolver_history_companion_rows": 56,
        "leverage_liquidity_resolved_rows": 48,
        "leverage_liquidity_availability_rows": 12,
        "leverage_liquidity_companion_rows": 36,
        "debt_credit_notes_resolved_rows": 16,
        "debt_maturity_ladder_resolved_rows": 0,
    }
    serialized_plan = value_plan.to_dict()
    assert serialized_plan["planned_write_count"] == 23_521
    assert serialized_plan["structured_skip_count"] == 2_012
    assert value_plan.issue_ledger["summary"]["canonical_unique_issue_count"] == 761
    assert value_plan.issue_ledger["summary"]["detailed_occurrence_count"] == 2_323
    assert len(style_plan.actions) == 770
    assert len(style_plan.decisions) == 1_298
    assert dict(value_plan.sheet_visibility)["Debt_Maturity_Ladder"] == "hidden"
    debt_policy_ids = {
        "debt_profile_product_state",
        "revolver_history_product_state",
        "leverage_liquidity_product_state",
        "debt_credit_notes_product_state",
        "debt_maturity_product_state",
    }
    debt_actions = [action for action in style_plan.actions if action.policy_id in debt_policy_ids]
    assert len(debt_actions) == 17
    assert {action.cell for action in debt_actions} <= {
        *(f"H{row}" for row in range(4, 15)),
        *(f"O{row}" for row in range(4, 16)),
        *(f"L{row}" for row in range(4, 16)),
    }


def test_checked_in_shell_has_exact_debt_topology_formulas_and_protection() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        expected = {
            "Debt_Profile": ("Debt and liquidity profile", "A4", 10),
            "Revolver_History": ("Revolver history", "A4", 16),
            "Leverage_Liquidity": ("Leverage and liquidity", "A4", 14),
            "Debt_Credit_Notes": ("Debt and credit notes", "A4", 8),
            "Debt_Maturity_Ladder": ("Debt maturity ladder", "A4", 8),
        }
        for sheet, (title, freeze, header_count) in expected.items():
            ws = wb[sheet]
            assert ws["A1"].value == title
            assert ws.freeze_panes == freeze
            assert sum(ws.cell(3, column).value is not None for column in range(1, header_count + 1)) == header_count
            assert ws.protection.sheet is True
            assert ws.sheet_state == "hidden"

        history = wb["Revolver_History"]
        liquidity = wb["Leverage_Liquidity"]
        assert history["K4"].value == '=IFERROR(IF(OR(NOT(ISNUMBER(F4)),NOT(ISNUMBER(D4)),D4=0),"",F4/D4),"")'
        assert liquidity["F4"].value == '=IFERROR(IF(OR(NOT(ISNUMBER(D4)),NOT(ISNUMBER(B4))),"",D4-B4),"")'
        assert liquidity["H4"].value == '=IFERROR(IF(OR(NOT(ISNUMBER(B4)),NOT(ISNUMBER(G4))),"",B4+G4),"")'
        assert "COUNTIF('Valuation'!$B$6:$M$6,$A4)<>1" in str(liquidity["K4"].value)
        for coordinate in ("F4", "H4", "I4", "J4", "K4"):
            assert liquidity[coordinate].protection.locked is True
    finally:
        wb.close()


def test_filled_anf_debt_product_is_exact_and_preserves_existing_snapshot(tmp_path: Path) -> None:
    output = tmp_path / "ANF_debt_product.xlsx"
    result = fill_standard_template_from_package(ANF_PACKAGE, output_path=output)

    assert result.written_cell_count == 23_521
    assert result.styled_cell_count == 770
    wb = load_workbook(output, data_only=False, read_only=False)
    try:
        assert {
            sheet: wb[sheet].sheet_state
            for sheet in (
                "Debt_Profile",
                "Revolver_History",
                "Leverage_Liquidity",
                "Debt_Credit_Notes",
                "Debt_Maturity_Ladder",
                "Debt_Tranches_Latest",
                "Debt_Tranches_Q",
                "Debt_Buckets",
                "Debt_Recon",
            )
        } == {
            "Debt_Profile": "visible",
            "Revolver_History": "visible",
            "Leverage_Liquidity": "visible",
            "Debt_Credit_Notes": "visible",
            "Debt_Maturity_Ladder": "hidden",
            "Debt_Tranches_Latest": "hidden",
            "Debt_Tranches_Q": "hidden",
            "Debt_Buckets": "hidden",
            "Debt_Recon": "hidden",
        }

        profile = wb["Debt_Profile"]
        assert [profile[f"D{row}"].value for row in range(4, 14)] == pytest.approx(
            [500.0, 500.0, 0.0, 0.469, 499.531, 50.0, 449.531, 594.08, 7.336, 1292.477]
        )
        assert profile["D14"].value is None
        assert profile["H6"].value == "reported_zero"
        assert profile["H14"].value == "unavailable"
        assert profile["J14"].value == "no source-backed core borrowings identified"
        assert all(profile[f"F{row}"].value == "2026-05-02" for row in range(4, 15))
        assert all(
            profile.cell(row, column).value is None
            for row in range(15, 17)
            for column in range(1, 11)
        )
        assert all(
            profile.cell(row, column).protection.locked
            for row in range(4, 17)
            for column in range(1, 11)
        )

        history = wb["Revolver_History"]
        assert [history[f"A{row}"].value for row in range(4, 16)] == [
            "2023-07-29",
            "2023-10-28",
            "2024-02-03",
            "2024-05-04",
            "2024-08-03",
            "2024-11-02",
            "2025-02-01",
            "2025-05-03",
            "2025-08-02",
            "2025-11-01",
            "2026-01-31",
            "2026-05-02",
        ]
        assert history["F8"].value is None and history["F9"].value is None
        assert history["O8"].value == history["O9"].value == "accepted_with_missing_drawn"
        assert history["F4"].value == 0
        assert str(history["K4"].value).startswith("=IFERROR(")

        liquidity = wb["Leverage_Liquidity"]
        rows_by_period = {liquidity[f"A{row}"].value: row for row in range(4, 16)}
        may_row = rows_by_period["2026-Q1"]
        january_row = rows_by_period["2025-Q4"]
        assert liquidity[f"B{may_row}"].value == pytest.approx(594.08)
        assert liquidity[f"G{may_row}"].value == pytest.approx(449.531)
        assert liquidity[f"B{january_row}"].value == pytest.approx(759.54)
        assert liquidity[f"G{january_row}"].value == pytest.approx(449.546)
        assert liquidity[f"C{may_row}"].value == pytest.approx(7.336)
        assert liquidity[f"D{may_row}"].value is None
        assert liquidity[f"E{may_row}"].value == pytest.approx(1292.477)
        assert liquidity[f"L{may_row}"].value == "debt_unavailable"
        for column in "FHIJK":
            assert str(liquidity[f"{column}{may_row}"].value).startswith("=IFERROR(")

        notes = wb["Debt_Credit_Notes"]
        assert [notes[f"A{row}"].value for row in range(4, 6)] == [
            "Covenant compliance",
            "Senior-notes redemption / no ABL borrowings",
        ]
        assert all(
            notes.cell(row, column).value is None
            for row in range(6, 10)
            for column in range(1, 9)
        )
        for sheet, target in (
            ("Debt_Maturity_Ladder", "A4:H20"),
            ("Debt_Tranches_Latest", "A2:J200"),
            ("Debt_Tranches_Q", "A2:J500"),
            ("Debt_Buckets", "A2:H200"),
            ("Debt_Recon", "A2:J200"),
        ):
            assert all(cell.value is None for row in wb[sheet][target] for cell in row)

        assert wb["SUMMARY"]["B44"].value == pytest.approx(449.546)
        assert wb["SUMMARY"]["D44"].value == "As of 2026-01-31 (stale)"
        assert wb["SUMMARY"]["B45"].value == pytest.approx(1209.086)
        assert wb["SUMMARY"]["D45"].value == "As of 2026-01-31 (stale)"
        assert wb["Valuation"]["M95"].value == pytest.approx(449.531)
        for row, value, period in (
            (124, 594.08, "2026-05-02"),
            (125, 449.546, "2026-01-31"),
            (126, 1209.086, "2026-01-31"),
            (127, 1292.477, "2026-05-02"),
        ):
            assert wb["Valuation"][f"B{row}"].value == pytest.approx(value)
            assert wb["Valuation"][f"D{row}"].value == period
            assert wb["Valuation"][f"E{row}"].value == "populated"
    finally:
        wb.close()
