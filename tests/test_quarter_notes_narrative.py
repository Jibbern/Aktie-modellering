from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from pbi_xbrl.excel_writer_quarter_narrative import (
    format_quarter_notes_period_header,
    parse_quarter_notes_period_header,
    _quarter_narrative_recent_history_periods,
    _quarter_narrative_records_for_context,
    _write_quarter_notes_ui_narrative_sheet,
)
from pbi_xbrl.excel_writer import (
    _quarter_notes_ui_snapshot_from_ws,
    _validate_quarter_notes_ui_export_snapshot,
    read_quarter_notes_ui_snapshot,
)
from tests.workbook_test_resources import delivered_workbook_path


WORKBOOK_DIR = Path(
    os.environ.get(
        "STOCK_MODEL_WORKBOOK_DIR",
        r"C:\Users\Jibbe\Aktier\StockModelData\outputs\Excel stock models",
    )
)
TICKERS = ("PBI", "GPRE", "ANF")


@pytest.mark.parametrize(
    ("header", "fiscal_period", "snapshot_key"),
    [
        ("2026-Q2 - Quarter Notes", "2026-Q2", "2026-06-30"),
        ("2026-Q2", "2026-Q2", "2026-06-30"),
        ("Q2 2026", "2026-Q2", "2026-06-30"),
        ("2026-06-30", "2026-Q2", "2026-06-30"),
    ],
)
def test_quarter_notes_period_header_contract_roundtrips_supported_forms(
    header: str,
    fiscal_period: str,
    snapshot_key: str,
) -> None:
    identity = parse_quarter_notes_period_header(header)
    assert identity is not None
    assert identity.fiscal_period == fiscal_period
    assert identity.snapshot_key == snapshot_key
    assert identity.event_id == f"quarter-notes-event:{fiscal_period}"
    assert parse_quarter_notes_period_header(format_quarter_notes_period_header(fiscal_period)) == identity


def test_quarter_notes_period_header_contract_rejects_malformed_or_ambiguous_headers() -> None:
    for header in ("", "Quarter Notes", "2026-Q5 - Quarter Notes", "2026-Q2 - Notes", "Q2"):
        assert parse_quarter_notes_period_header(header) is None


def test_quarter_notes_saved_snapshot_reads_current_semantic_rows_and_fails_closed_on_empty_parse() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quarter_Notes_UI"
    ws["A1"] = "2026-Q2 - Quarter Notes"
    ws["A2"] = "Quarter read"
    ws["A8"] = "Key developments"
    ws["A9"] = "Theme"
    ws["C9"] = "What happened"
    ws["A10"] = "Source-backed milestone"
    ws["C10"] = "The milestone was completed in Q2."

    snapshot = _quarter_notes_ui_snapshot_from_ws(ws)
    assert snapshot == {
        "2026-06-30": [("Source-backed milestone", "The milestone was completed in Q2.")]
    }
    _validate_quarter_notes_ui_export_snapshot(snapshot, snapshot, Path("current.xlsx"))

    malformed = Workbook().active
    malformed.title = "Quarter_Notes_UI"
    malformed["A1"] = "2026-Q2 - Notes"
    malformed["B2"] = "Programs / initiatives"
    malformed["C2"] = "Visible content whose period header cannot be parsed."
    malformed_snapshot = _quarter_notes_ui_snapshot_from_ws(malformed)
    assert malformed_snapshot == {}
    with pytest.raises(RuntimeError, match="zero semantic rows"):
        _validate_quarter_notes_ui_export_snapshot({}, {}, Path("malformed.xlsx"))
    _validate_quarter_notes_ui_export_snapshot(
        {},
        {},
        Path("intentionally-empty.xlsx"),
        intentionally_empty=True,
    )


def test_current_delivered_quarter_notes_surfaces_parse_nonempty_for_every_profile() -> None:
    for ticker in TICKERS:
        snapshot = read_quarter_notes_ui_snapshot(delivered_workbook_path(ticker, Path(__file__)))
        assert snapshot
        assert any(rows for rows in snapshot.values())

REQUIRED_NARRATIVE_HEADERS = [
    "Ticker",
    "Quarter",
    "Category",
    "Theme",
    "What happened",
    "Management framing",
    "Why it matters",
    "Model implication",
    "Valuation implication",
    "Double-count guardrail",
    "Linked sheet",
    "Linked metric",
    "Amount",
    "Unit",
    "Source date",
    "Source type",
    "Source / note",
    "Confidence",
    "Include in UI",
]

NOISY_PATTERNS = [
    r"\[UPDATED\]",
    r"\bDEBUG\b",
    r"\bTODO\b",
    r"\bFIXME\b",
    r"raw_json",
    r"metadata_candidate",
    r"source_txt_file",
    r"source_txt",
    r"<\s*/?\s*(html|body|table|xml|xbrl|ix:)",
    r"\bSEC-HEADER\b",
    r"\b(operator|analyst|question-and-answer session)\s*:",
    r"\b(nan|null)\b",
]

SECTION_TITLES = {
    "Quarter read",
    "Key developments",
    "Guidance / Promise interpretation",
    "Model mapping / double-count guardrails",
}
TABLE_HEADER_LABELS = {
    "Theme",
    "Promise / guidance item",
    "Driver",
}


def _load_workbook(ticker: str, *, data_only: bool = True):
    path = next(
        (WORKBOOK_DIR / f"{ticker}_model{suffix}" for suffix in (".xlsm", ".xlsx") if (WORKBOOK_DIR / f"{ticker}_model{suffix}").exists()),
        WORKBOOK_DIR / f"{ticker}_model.xlsx",
    )
    if not path.exists():
        pytest.skip(f"{path} is not available for quarter narrative regression tests")
    return load_workbook(path, data_only=data_only, read_only=False)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _all_cell_text(ws: Worksheet) -> str:
    return "\n".join(
        _text(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if _text(cell.value)
    )


def _headers(ws: Worksheet) -> List[str]:
    return [_text(ws.cell(1, cc).value) for cc in range(1, int(ws.max_column or 0) + 1)]


def _narrative_rows(ws: Worksheet) -> List[Tuple[int, Dict[str, Any]]]:
    headers = _headers(ws)
    rows: List[Tuple[int, Dict[str, Any]]] = []
    for rr in range(2, int(ws.max_row or 0) + 1):
        row = {headers[cc - 1]: ws.cell(rr, cc).value for cc in range(1, len(headers) + 1)}
        if any(_text(value) for value in row.values()):
            rows.append((rr, row))
    return rows


def _include_in_ui(value: Any) -> bool:
    return _text(value).lower() in {"yes", "true", "1", "y"}


def _assert_text_has(text: str, terms: Sequence[str], *, context: str) -> None:
    low = text.lower()
    missing = [term for term in terms if term.lower() not in low]
    assert not missing, f"{context}: missing expected narrative terms {missing}"


def _effective_cell(ws: Worksheet, row: int, col: int) -> Cell:
    coord = ws.cell(row, col).coordinate
    for merged in ws.merged_cells.ranges:
        if coord in merged:
            return ws.cell(merged.min_row, merged.min_col)
    return ws.cell(row, col)


def _effective_fill(ws: Worksheet, row: int, col: int) -> str:
    cell = _effective_cell(ws, row, col)
    return _text(cell.fill.fgColor.rgb).upper()


def _effective_border_styles(ws: Worksheet, row: int, col: int) -> Tuple[str, str, str, str]:
    cell = _effective_cell(ws, row, col)
    return (
        _text(cell.border.left.style),
        _text(cell.border.right.style),
        _text(cell.border.top.style),
        _text(cell.border.bottom.style),
    )


def _effective_wrap(ws: Worksheet, row: int, col: int) -> bool:
    return bool(_effective_cell(ws, row, col).alignment.wrap_text)


def _is_merged_across_a_j(ws: Worksheet, row: int) -> bool:
    return any(
        merged.min_row == row
        and merged.max_row == row
        and merged.min_col == 1
        and merged.max_col >= 10
        for merged in ws.merged_cells.ranges
    )


def _quarter_header_rows(ws: Worksheet) -> List[int]:
    return [
        rr
        for rr in range(1, int(ws.max_row or 0) + 1)
        if _text(ws.cell(rr, 1).value).endswith(" - Quarter Notes")
    ]


def _key_development_rows(ws: Worksheet) -> List[Tuple[int, str, str, str]]:
    rows: List[Tuple[int, str, str, str]] = []
    rr = 1
    while rr <= int(ws.max_row or 0):
        if _text(ws.cell(rr, 1).value) != "Key developments":
            rr += 1
            continue
        body = rr + 2
        while body <= int(ws.max_row or 0):
            first = _text(ws.cell(body, 1).value)
            if not first or first in SECTION_TITLES or first.endswith(" - Quarter Notes"):
                break
            if first not in TABLE_HEADER_LABELS:
                source = next(
                    (_text(ws.cell(body, cc).value) for cc in range(13, min(15, int(ws.max_column or 0)) + 1) if _text(ws.cell(body, cc).value)),
                    "",
                ) or _text(ws.cell(body, 11).value) or _text(ws.cell(body, 10).value)
                rows.append((body, first, _text(ws.cell(body, 3).value), source))
            body += 1
        rr = body
    return rows


def _model_mapping_rows(ws: Worksheet) -> List[Tuple[int, str, str, str, str]]:
    rows: List[Tuple[int, str, str, str, str]] = []
    rr = 1
    while rr <= int(ws.max_row or 0):
        if _text(ws.cell(rr, 1).value) != "Model mapping / double-count guardrails":
            rr += 1
            continue
        body = rr + 2
        while body <= int(ws.max_row or 0):
            first = _text(ws.cell(body, 1).value)
            if not first or first in SECTION_TITLES or first.endswith(" - Quarter Notes"):
                break
            if first != "Driver":
                linked = next(
                    (_text(ws.cell(body, cc).value) for cc in range(13, min(15, int(ws.max_column or 0)) + 1) if _text(ws.cell(body, cc).value)),
                    "",
                ) or _text(ws.cell(body, 10).value) or _text(ws.cell(body, 9).value)
                rows.append(
                    (
                        body,
                        first,
                        _text(ws.cell(body, 3).value),
                        _text(ws.cell(body, 7).value) or _text(ws.cell(body, 6).value),
                        linked,
                    )
                )
            body += 1
        rr = body
    return rows


def _promise_interpretation_rows(ws: Worksheet) -> List[Tuple[int, str, str, str, str]]:
    rows: List[Tuple[int, str, str, str, str]] = []
    rr = 1
    while rr <= int(ws.max_row or 0):
        if _text(ws.cell(rr, 1).value) != "Guidance / Promise interpretation":
            rr += 1
            continue
        body = rr + 2
        while body <= int(ws.max_row or 0):
            first = _text(ws.cell(body, 1).value)
            if not first or first in SECTION_TITLES or first.endswith(" - Quarter Notes"):
                break
            if first != "Promise / guidance item":
                source = next(
                    (_text(ws.cell(body, cc).value) for cc in range(13, min(15, int(ws.max_column or 0)) + 1) if _text(ws.cell(body, cc).value)),
                    "",
                ) or _text(ws.cell(body, 11).value) or _text(ws.cell(body, 10).value)
                rows.append(
                    (
                        body,
                        first,
                        _text(ws.cell(body, 3).value),
                        _text(ws.cell(body, 9).value) or _text(ws.cell(body, 8).value),
                        source,
                    )
                )
            body += 1
        rr = body
    return rows


def _assert_no_noisy_text(ticker: str, sheet_name: str, ws: Worksheet) -> None:
    compiled = [(pattern, re.compile(pattern, re.I)) for pattern in NOISY_PATTERNS]
    hits: List[str] = []
    long_hits: List[str] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            txt = value.strip()
            if len(txt) > 600:
                long_hits.append(f"{cell.coordinate} len={len(txt)}")
            for pattern, regex in compiled:
                if regex.search(txt):
                    hits.append(f"{cell.coordinate}: {pattern!r} in {txt[:120]!r}")
    assert not hits, f"{ticker} {sheet_name}: noisy/debug text found: {hits[:10]}"
    assert not long_hits, f"{ticker} {sheet_name}: suspiciously long raw-like text found: {long_hits[:10]}"


def test_quarter_narrative_data_has_required_structure_and_ui_traceability() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        try:
            assert "Quarter_Narrative_Data" in wb.sheetnames, f"{ticker}: missing Quarter_Narrative_Data"
            assert "Quarter_Notes_UI" in wb.sheetnames, f"{ticker}: missing Quarter_Notes_UI"
            data_ws = wb["Quarter_Narrative_Data"]
            ui_ws = wb["Quarter_Notes_UI"]

            assert _headers(data_ws) == REQUIRED_NARRATIVE_HEADERS, f"{ticker}: Quarter_Narrative_Data headers changed"
            rows = _narrative_rows(data_ws)
            assert rows, f"{ticker}: no narrative rows"
            for rr, row in rows:
                assert _text(row["Ticker"]) == ticker, f"{ticker}: row {rr} has wrong ticker"
                assert _text(row["Quarter"]), f"{ticker}: row {rr} missing Quarter"
                assert _text(row["Category"]), f"{ticker}: row {rr} missing Category"
                assert _text(row["Theme"]), f"{ticker}: row {rr} missing Theme"
                if _include_in_ui(row["Include in UI"]):
                    useful_fields = [
                        "What happened",
                        "Why it matters",
                        "Model implication",
                        "Valuation implication",
                        "Double-count guardrail",
                    ]
                    assert any(_text(row[field]) for field in useful_fields), f"{ticker}: UI row {rr} has no useful narrative"
                    assert _text(row["Source date"]) or _text(row["Source / note"]), f"{ticker}: UI row {rr} lacks source context"
                    assert _text(row["Source type"]) or _text(row["Source / note"]), f"{ticker}: UI row {rr} lacks source type/note"
                    assert _text(row["Confidence"]).lower() in {"high", "medium", "low"}, f"{ticker}: row {rr} lacks confidence"
                    if _text(row["Confidence"]).lower() == "low":
                        visible_blob = " ".join(_text(row[field]) for field in useful_fields)
                        assert "interpret" in visible_blob.lower() or "review" in visible_blob.lower(), f"{ticker}: low confidence row {rr} needs visible caveat"
                if _text(row["Amount"]) or _text(row["Unit"]):
                    assert _text(row["Source date"]) and (_text(row["Source type"]) or _text(row["Source / note"])), (
                        f"{ticker}: amount/unit row {rr} lacks source context"
                    )

            data_themes = {_text(row["Theme"]) for _, row in rows if _include_in_ui(row["Include in UI"])}
            for rr, theme, what_happened, source in _key_development_rows(ui_ws):
                assert theme in data_themes, f"{ticker}: UI key development row {rr} theme {theme!r} not traceable to Quarter_Narrative_Data"
                assert what_happened, f"{ticker}: UI key development row {rr} missing narrative"
                assert source and re.search(r"\b(high|medium|low)\b", source, re.I), f"{ticker}: UI key development row {rr} lacks source/confidence"
        finally:
            wb.close()


def test_quarter_narrative_and_ui_have_no_noisy_raw_text() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        try:
            _assert_no_noisy_text(ticker, "Quarter_Narrative_Data", wb["Quarter_Narrative_Data"])
            _assert_no_noisy_text(ticker, "Quarter_Notes_UI", wb["Quarter_Notes_UI"])
        finally:
            wb.close()


def test_quarter_narrative_amount_excludes_descriptor_prose() -> None:
    descriptor_patterns = (
        "are geographic segments",
        "are brand",
    )
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        try:
            data_ws = wb["Quarter_Narrative_Data"]
            headers = _headers(data_ws)
            amount_col = headers.index("Amount") + 1
            rows = _narrative_rows(data_ws)
            descriptor_hits: List[str] = []
            prose_hits: List[str] = []
            for rr, row in rows:
                amount_txt = _text(data_ws.cell(rr, amount_col).value)
                if not amount_txt:
                    continue
                amount_low = amount_txt.lower()
                if any(pattern in amount_low for pattern in descriptor_patterns):
                    descriptor_hits.append(f"{ticker} Quarter_Narrative_Data!M{rr}: {amount_txt}")
                if len(amount_txt) > 55 and not re.search(r"[$%0-9]", amount_txt):
                    prose_hits.append(f"{ticker} Quarter_Narrative_Data!M{rr}: {amount_txt}")
            assert not descriptor_hits, "Amount column should not contain segment descriptor prose: " + "; ".join(descriptor_hits[:5])
            assert not prose_hits, "Amount column should remain value-like, not long prose: " + "; ".join(prose_hits[:5])

            if ticker == "ANF":
                anf_segment_rows = [
                    row
                    for _rr, row in rows
                    if _text(row.get("Quarter")) == "2025-Q4"
                    and re.search(
                        r"\b(americas|emea|apac|abercrombie|hollister)\b",
                        f"{_text(row.get('Theme'))} {_text(row.get('Linked metric'))}".lower(),
                    )
                    and re.search(r"[$%0-9]", _text(row.get("Amount")))
                ]
                assert anf_segment_rows, "ANF 2025-Q4 should retain numeric brand/geography narrative rows"
        finally:
            wb.close()


def test_quarter_notes_cover_ticker_specific_narratives_and_guardrails() -> None:
    expectations = {
        "PBI": {
            "combined_terms": [
                "GEC",
                "$136m",
                "cost reductions",
                "$70m",
                "$120m-$160m",
                "cash optimization",
                "$240m",
                "FCF",
                "Adjusted FCF",
                "Presort",
                "SendTech",
                "illustrative EBIT",
                "forecast",
            ],
            "guardrail_terms": ["GEC", "cost savings", "illustrative EBIT bridge", "separate"],
        },
        "GPRE": {
            "combined_terms": [
                "45Z",
                "monetized",
                "$23.4m",
                "$200m-$225m",
                "facility qualification",
                "operational",
                "Advantage Nebraska",
                "crush",
                "RVO",
                "E15",
                "export",
                "capex",
                "FCF",
                "liquidity",
            ],
            "guardrail_terms": ["incremental", "baseline", "full guidance"],
        },
        "ANF": {
            "combined_terms": [
                "sales growth",
                "Abercrombie",
                "Hollister",
                "operating margin",
                "tariff",
                "freight",
                "ERP",
                "marketing",
                "buybacks",
                "share count",
                "capex",
                "pre-release",
                "Adjusted EPS",
                "GAAP EPS",
                "bps",
            ],
            "ui_terms": ["Americas", "EMEA", "APAC"],
            "guardrail_terms": ["bps", "operating margin", "double-count"],
        },
    }

    for ticker, expected in expectations.items():
        wb = _load_workbook(ticker)
        try:
            data_ws = wb["Quarter_Narrative_Data"]
            ui_ws = wb["Quarter_Notes_UI"]
            combined_text = _all_cell_text(data_ws) + "\n" + _all_cell_text(ui_ws)
            ui_text = _all_cell_text(ui_ws)

            _assert_text_has(combined_text, expected["combined_terms"], context=f"{ticker} combined narrative")
            _assert_text_has(ui_text, expected.get("ui_terms", []), context=f"{ticker} Quarter_Notes_UI")
            _assert_text_has(combined_text, expected["guardrail_terms"], context=f"{ticker} double-count guardrails")

            rows = [row for _, row in _narrative_rows(data_ws)]
            themes = {_text(row["Theme"]) for row in rows}
            if ticker == "PBI":
                assert {"GEC loss removal", "Annual savings target", "Implemented cost reductions", "Cash optimization", "Illustrative EBIT bridge"}.issubset(themes)
                gec_rows = [row for row in rows if _text(row["Theme"]) == "GEC loss removal"]
                assert gec_rows and "cost savings" in _text(gec_rows[0]["Double-count guardrail"]).lower()
            if ticker == "GPRE":
                qualification_rows = [row for row in rows if "facility qualification" in _text(row["Theme"]).lower()]
                assert qualification_rows, "GPRE: missing 45Z facility qualification narrative"
                qualification_blob = " ".join(_text(value) for value in qualification_rows[0].values())
                _assert_text_has(qualification_blob, ["operational", "progress", "not as completed"], context="GPRE qualification evidence")
            if ticker == "ANF":
                assert {"Tariff headwind", "Freight tailwind", "ERP disruption", "Marketing headwind", "Brand and geography cuts"}.issubset(themes)
        finally:
            wb.close()


def test_quarter_narrative_data_dedupes_sendtech_theme_within_quarter() -> None:
    wb = _load_workbook("PBI")
    try:
        rows = _narrative_rows(wb["Quarter_Narrative_Data"])
        counts: Dict[Tuple[str, str], int] = {}
        examples: Dict[Tuple[str, str], List[int]] = {}
        for rr, row in rows:
            blob = " ".join(
                _text(row.get(field))
                for field in ("Theme", "What happened", "Linked metric", "Model implication")
            ).lower()
            if "sendtech" not in blob:
                continue
            quarter = _text(row.get("Quarter"))
            theme_key = re.sub(r"\s+", " ", _text(row.get("Theme")).lower()).strip()
            key = (quarter, theme_key)
            counts[key] = counts.get(key, 0) + 1
            examples.setdefault(key, []).append(rr)
        duplicates = {key: rr_list for key, rr_list in examples.items() if counts.get(key, 0) > 1}
        assert not duplicates, f"PBI Quarter_Narrative_Data duplicate SendTech themes within quarter: {duplicates}"
    finally:
        wb.close()


def test_quarter_notes_layout_is_readable_and_continuous_across_a_to_j() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        try:
            ws = wb["Quarter_Notes_UI"]
            quarter_rows = _quarter_header_rows(ws)
            assert quarter_rows, f"{ticker}: no quarter headers in Quarter_Notes_UI"
            for rr in quarter_rows:
                assert _is_merged_across_a_j(ws, rr), f"{ticker}: quarter header row {rr} is not merged A:J"
                assert _effective_fill(ws, rr, 1).endswith(("5B9BD5", "4472C4")), f"{ticker}: quarter header row {rr} missing blue fill"

            for rr in range(1, int(ws.max_row or 0) + 1):
                height = ws.row_dimensions[rr].height or 0
                assert height <= 90, f"{ticker}: row {rr} height {height} is a readability blowup"
                row_values = [_text(ws.cell(rr, cc).value) for cc in range(1, 13)]
                if not any(row_values):
                    continue
                first = row_values[0]
                if first.endswith(" - Quarter Notes"):
                    continue
                if first in SECTION_TITLES or first in TABLE_HEADER_LABELS or first in {"Model read", "What changed", "Watch next", "Key caveat"}:
                    assert _effective_fill(ws, rr, 1) not in {"", "00000000"}, f"{ticker}: row {rr} missing section/body fill"
                if first not in SECTION_TITLES and first not in TABLE_HEADER_LABELS:
                    for cc in range(1, 13):
                        assert _effective_fill(ws, rr, cc) not in {"", "00000000"}, f"{ticker}: row {rr} col {cc} has zebra fill gap"
                        assert any(_effective_border_styles(ws, rr, cc)), f"{ticker}: row {rr} col {cc} missing border"
                    if any(row_values[1:]):
                        assert any(_effective_wrap(ws, rr, cc) for cc in range(2, 13)), f"{ticker}: row {rr} narrative cells are not wrapped"

            key_rows = _key_development_rows(ws)
            promise_rows = _promise_interpretation_rows(ws)
            mapping_rows = _model_mapping_rows(ws)
            assert key_rows, f"{ticker}: missing key development rows"
            assert mapping_rows, f"{ticker}: missing model mapping guardrail rows"
            assert promise_rows or ticker == "ANF", f"{ticker}: expected promise interpretation rows where guidance is present"

            previous: Tuple[str, str] | None = None
            duplicates: List[int] = []
            for rr, theme, what_happened, _source in key_rows:
                current = (theme, what_happened)
                if current == previous:
                    duplicates.append(rr)
                previous = current
            assert not duplicates, f"{ticker}: duplicate consecutive key development rows at {duplicates}"

            for rr, driver, treatment, guardrail, linked in mapping_rows:
                assert treatment, f"{ticker}: model mapping row {rr} missing model treatment"
                assert guardrail, f"{ticker}: model mapping row {rr} missing double-count guardrail"
                assert linked and "|" in linked, f"{ticker}: model mapping row {rr} lacks linked sheet/metric"
                assert driver.lower() not in {"metric", "driver"}, f"{ticker}: stub mapping row {rr}"
        finally:
            wb.close()


def test_quarter_notes_references_are_consistent_with_promise_and_investment_case() -> None:
    concept_sheet_tokens = {
        "Investment_Case",
        "Scenario Driver Bridge",
        "Quarter_Notes_UI",
        "Scenario_Driver_Assumptions",
        "Scenario_Bridge_Tax_Treatment",
        "Operating_Drivers",
        "Valuation",
        "BS_Segments",
        "Debt_Profile",
        "Promise_Progress_UI",
    }
    critical_investment_case_refs = {
        "PBI": ["Incremental cost savings vs baseline", "Segment Scenario Inputs"],
        "GPRE": ["Incremental 45Z uplift vs baseline", "Capex change vs baseline"],
        "ANF": ["Margin bridge vs baseline", "Tariff impact (bps)", "Freight tailwind (bps)", "ERP disruption (bps)", "Marketing headwind (bps)"],
    }

    for ticker in TICKERS:
        wb = _load_workbook(ticker, data_only=False)
        try:
            periods = _quarter_narrative_recent_history_periods(wb, limit=8)
            records = _quarter_narrative_records_for_context(
                ticker,
                workbook=wb,
                history_periods=periods,
                max_per_period=5,
            )
            assert _write_quarter_notes_ui_narrative_sheet(
                wb,
                ticker,
                records,
                history_periods=periods,
            )
            data_ws = wb["Quarter_Narrative_Data"]
            ui_ws = wb["Quarter_Notes_UI"]
            promise_text = _all_cell_text(wb["Promise_Progress_UI"])
            ic_text = _all_cell_text(wb[f"{ticker}_Investment_Case"])
            ui_text = _all_cell_text(ui_ws)

            for _, row in _narrative_rows(data_ws):
                linked_sheet_blob = _text(row["Linked sheet"])
                if not linked_sheet_blob:
                    continue
                for token in [part.strip() for part in linked_sheet_blob.split(";") if part.strip()]:
                    assert token in concept_sheet_tokens or token in wb.sheetnames or token == f"{ticker}_Investment_Case", (
                        f"{ticker}: unknown linked sheet token {token!r}"
                    )

            for phrase in critical_investment_case_refs[ticker]:
                assert phrase in ic_text, f"{ticker}: expected Investment_Case reference {phrase!r}"
                assert phrase in ui_text or phrase.replace("Incremental ", "") in ui_text or phrase.split(" (")[0] in ui_text, (
                    f"{ticker}: Quarter_Notes_UI does not explain Investment_Case driver {phrase!r}"
                )

            for rr, item, read, caveat, source in _promise_interpretation_rows(ui_ws):
                assert source, f"{ticker}: promise interpretation row {rr} missing source"
                item_key = item.split(" / ")[0].split(" (")[0]
                official_source = any(token in source.lower() for token in ("earnings release", "presentation", "pre-release", "operating_drivers"))
                assert item_key.lower() in promise_text.lower() or official_source or "source-backed" in source.lower() or "Investment_Case" in source, (
                    f"{ticker}: promise interpretation row {rr} has no Promise_Progress_UI/source-backed anchor"
                )
                assert read or caveat, f"{ticker}: promise interpretation row {rr} lacks interpretation"
        finally:
            wb.close()
