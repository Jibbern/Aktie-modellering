from __future__ import annotations

import os
from pathlib import Path
import sys

import pytest
from openpyxl import Workbook, load_workbook

import pbi_xbrl.render_validation_runner as render_runner
from pbi_xbrl.render_validation_runner import (
    RENDER_RANGES,
    RenderTarget,
    USER_FACING_STYLE_SHEETS,
    _default_workbooks,
    _render_ranges_with_excel_com,
    discover_render_targets,
    run_render_validation,
    validate_openpyxl_layout,
)


WORKBOOK_DIR = Path(
    os.environ.get(
        "STOCK_MODEL_WORKBOOK_DIR",
        r"C:\Users\Jibbe\Aktier\StockModelData\outputs\Excel stock models",
    )
)
TICKERS = ("PBI", "GPRE", "ANF")


def _workbook_path(ticker: str) -> Path:
    path = next(
        (WORKBOOK_DIR / f"{ticker}_model{suffix}" for suffix in (".xlsm", ".xlsx") if (WORKBOOK_DIR / f"{ticker}_model{suffix}").exists()),
        WORKBOOK_DIR / f"{ticker}_model.xlsx",
    )
    if not path.exists():
        pytest.skip(f"{path} is not available for render/style validation tests")
    return path


def _legacy_core_render_targets(ticker: str) -> tuple[RenderTarget, ...]:
    return tuple(
        RenderTarget(
            sheet=sheet_template.format(ticker=ticker),
            range_ref=range_ref,
            source="core_render_contract",
        )
        for sheet_template, range_ref in RENDER_RANGES.items()
    )


def test_render_ranges_cover_required_workbook_surfaces() -> None:
    expected_ranges = {
        "Valuation": "A1:AC90",
        "{ticker}_Investment_Case": "A1:J160",
        "Promise_Progress_UI": "A1:L180",
        "Quarter_Notes_UI": "A1:O220",
        "Operating_Drivers": "A1:Q140",
        "Needs_Review": "A1:J80",
    }

    assert RENDER_RANGES == expected_ranges
    assert {"Valuation", "{ticker}_Investment_Case", "Promise_Progress_UI", "Quarter_Notes_UI"}.issubset(
        set(USER_FACING_STYLE_SHEETS)
    )


def test_render_runner_prefers_generated_xlsm_workbooks(tmp_path: Path) -> None:
    (tmp_path / "PBI_model.xlsm").write_bytes(b"placeholder")
    (tmp_path / "PBI_model.xlsx").write_bytes(b"stale-placeholder")
    (tmp_path / "GPRE_model.xlsm").write_bytes(b"placeholder")

    workbooks = _default_workbooks(tmp_path)

    assert workbooks["PBI"] == tmp_path / "PBI_model.xlsm"
    assert workbooks["GPRE"] == tmp_path / "GPRE_model.xlsm"
    assert workbooks["ANF"] == tmp_path / "ANF_model.xlsx"


def test_render_target_discovery_adds_only_visible_manifest_product_blocks(tmp_path: Path) -> None:
    wb = Workbook()
    wb.active.title = "Valuation"
    for name in (
        "ANF_Investment_Case",
        "Promise_Progress_UI",
        "Quarter_Notes_UI",
        "Operating_Drivers",
        "Needs_Review",
        "Debt_Profile",
        "Revolver_History",
        "Leverage_Liquidity",
        "Debt_Credit_Notes",
        "Debt_Maturity_Ladder",
        "Debt_Tranches_Latest",
    ):
        wb.create_sheet(name)
    for name in wb.sheetnames:
        wb[name]["A1"] = name
    for name in ("Debt_Profile", "Revolver_History", "Leverage_Liquidity", "Debt_Credit_Notes"):
        wb[name].sheet_state = "visible"
        wb[name]["A4"] = "source-backed row"
    wb["Debt_Maturity_Ladder"].sheet_state = "hidden"
    wb["Debt_Tranches_Latest"].sheet_state = "hidden"
    path = tmp_path / "ANF_model.xlsx"
    wb.save(path)
    wb.close()

    targets = discover_render_targets(path, "ANF")

    assert [(target.sheet, target.range_ref) for target in targets[-4:]] == [
        ("Debt_Profile", "A1:J16"),
        ("Revolver_History", "A1:P15"),
        ("Leverage_Liquidity", "A1:N15"),
        ("Debt_Credit_Notes", "A1:H9"),
    ]
    assert len(targets) == 10
    assert all(target.sheet not in {"Debt_Maturity_Ladder", "Debt_Tranches_Latest"} for target in targets)


class _FakeRange:
    Left = 0.0
    Top = 0.0
    Width = 240.0
    Height = 120.0

    def __init__(self, *, copy_error: BaseException | None = None) -> None:
        self.copy_error = copy_error

    def Select(self) -> None:  # noqa: N802 - COM spelling
        return None

    def CopyPicture(self, **_kwargs: object) -> None:  # noqa: N802 - COM spelling
        if self.copy_error is not None:
            raise self.copy_error


class _FakeChart:
    def Paste(self) -> None:  # noqa: N802 - COM spelling
        return None

    def Export(self, path: str) -> bool:  # noqa: N802 - COM spelling
        Path(path).write_bytes(b"nonblank-render-fixture" * 20)
        return True


class _FakeChartObject:
    def __init__(self, *, delete_error: BaseException | None = None) -> None:
        self.Chart = _FakeChart()
        self.deleted = False
        self.delete_error = delete_error

    def Delete(self) -> None:  # noqa: N802 - COM spelling
        if self.delete_error is not None:
            raise self.delete_error
        self.deleted = True


class _FakeChartObjects:
    def __init__(self, *, allowed: bool, delete_error: BaseException | None = None) -> None:
        self.allowed = allowed
        self.delete_error = delete_error

    def Add(self, *_args: object) -> _FakeChartObject:  # noqa: N802 - COM spelling
        if not self.allowed:
            raise AssertionError("The source worksheet must never own a render chart.")
        return _FakeChartObject(delete_error=self.delete_error)


class _FakeWorksheet:
    Visible = -1

    def __init__(
        self,
        *,
        scratch: bool,
        copy_error: BaseException | None = None,
        chart_delete_error: BaseException | None = None,
    ) -> None:
        self.scratch = scratch
        self.copy_error = copy_error
        self.chart_delete_error = chart_delete_error

    def Activate(self) -> None:  # noqa: N802 - COM spelling
        return None

    def Unprotect(self) -> None:  # noqa: N802 - COM spelling
        return None

    def Range(self, _range_ref: str) -> _FakeRange:  # noqa: N802 - COM spelling
        return _FakeRange(copy_error=self.copy_error)

    def ChartObjects(self) -> _FakeChartObjects:  # noqa: N802 - COM spelling
        return _FakeChartObjects(
            allowed=self.scratch,
            delete_error=self.chart_delete_error,
        )


class _FakeWorkbook:
    def __init__(
        self,
        *,
        scratch: bool,
        copy_error: BaseException | None = None,
        chart_delete_error: BaseException | None = None,
        close_error: BaseException | None = None,
    ) -> None:
        self.sheet = _FakeWorksheet(
            scratch=scratch,
            copy_error=copy_error,
            chart_delete_error=chart_delete_error,
        )
        self.closed = False
        self.close_error = close_error

    def Activate(self) -> None:  # noqa: N802 - COM spelling
        return None

    def Worksheets(self, _key: object) -> _FakeWorksheet:  # noqa: N802 - COM spelling
        return self.sheet

    def Close(self, **_kwargs: object) -> None:  # noqa: N802 - COM spelling
        if self.close_error is not None:
            raise self.close_error
        self.closed = True


class _FakeWorkbooks:
    def __init__(
        self,
        *,
        copy_error: BaseException | None = None,
        chart_delete_error: BaseException | None = None,
        scratch_close_error: BaseException | None = None,
    ) -> None:
        self.scratch = _FakeWorkbook(
            scratch=True,
            chart_delete_error=chart_delete_error,
            close_error=scratch_close_error,
        )
        self.source = _FakeWorkbook(scratch=False, copy_error=copy_error)

    def Add(self) -> _FakeWorkbook:  # noqa: N802 - COM spelling
        return self.scratch

    def Open(self, *_args: object, **_kwargs: object) -> _FakeWorkbook:  # noqa: N802 - COM spelling
        return self.source


class _FakeExcel:
    Hwnd = 101

    def __init__(
        self,
        *,
        copy_error: BaseException | None = None,
        chart_delete_error: BaseException | None = None,
        scratch_close_error: BaseException | None = None,
        quit_error: BaseException | None = None,
    ) -> None:
        self.Workbooks = _FakeWorkbooks(
            copy_error=copy_error,
            chart_delete_error=chart_delete_error,
            scratch_close_error=scratch_close_error,
        )
        self.CutCopyMode = False
        self.quit_error = quit_error

    def Quit(self) -> None:  # noqa: N802 - COM spelling
        if self.quit_error is not None:
            raise self.quit_error
        return None


def _install_fake_excel(monkeypatch: pytest.MonkeyPatch, excel: _FakeExcel) -> None:
    monkeypatch.setattr(render_runner, "_co_initialize", lambda: None)
    monkeypatch.setattr(render_runner, "_co_uninitialize", lambda: None)
    monkeypatch.setattr(render_runner, "_dispatch_excel", lambda: excel)
    monkeypatch.setattr(render_runner, "_excel_process_id", lambda _excel: 1234)
    monkeypatch.setattr(
        render_runner,
        "_wait_for_owned_process_exit",
        lambda _process_id: {"status": "PASS", "forced_termination": False},
    )


def test_com_renderer_uses_scratch_chart_and_records_complete_telemetry(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"immutable source workbook")
    _install_fake_excel(monkeypatch, _FakeExcel())
    targets = {"ANF": (RenderTarget("Protected", "A1:C5", "test"),)}

    results = _render_ranges_with_excel_com({"ANF": source}, tmp_path / "images", targets)

    assert len(results) == 1
    result = results[0]
    assert result.status == "pass"
    assert result.source_bytes_unchanged is True
    assert source.read_bytes() == b"immutable source workbook"
    assert [row["operation"] for row in result.operations] == list(render_runner._RENDER_OPERATION_NAMES)
    assert all(row["status"] == "pass" for row in result.operations)
    assert result.owned_excel_process_cleanup == "PASS"


def test_com_renderer_preserves_hresult_and_exact_copy_picture_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class CopyPictureError(RuntimeError):
        hresult = -2146827284

    source = tmp_path / "source.xlsx"
    source.write_bytes(b"immutable source workbook")
    _install_fake_excel(monkeypatch, _FakeExcel(copy_error=CopyPictureError("CopyPicture failed")))
    monkeypatch.setattr(render_runner.time, "sleep", lambda _seconds: None)
    targets = {"ANF": (RenderTarget("Protected", "A1:C5", "test"),)}

    result = _render_ranges_with_excel_com({"ANF": source}, tmp_path / "images", targets)[0]

    copy_picture = next(row for row in result.operations if row["operation"] == "copy_picture")
    assert result.status == "fail"
    assert result.failing_operation == "copy_picture"
    assert copy_picture["attempts"] == 4
    assert copy_picture["hresult"] == -2146827284
    assert not list((tmp_path / "images").rglob("*.png"))


def test_com_renderer_reports_cleanup_only_failure_as_primary_cleanup(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class CleanupError(RuntimeError):
        hresult = -2147000001

    source = tmp_path / "source.xlsx"
    source.write_bytes(b"immutable source workbook")
    _install_fake_excel(
        monkeypatch,
        _FakeExcel(chart_delete_error=CleanupError("chart delete failed")),
    )
    targets = {"ANF": (RenderTarget("Protected", "A1:C5", "test"),)}

    result = _render_ranges_with_excel_com({"ANF": source}, tmp_path / "images", targets)[0]

    cleanup = next(row for row in result.operations if row["operation"] == "cleanup")
    assert result.status == "fail"
    assert result.failing_operation == "cleanup"
    assert cleanup["status"] == "fail"
    assert cleanup["hresult"] == -2147000001
    assert result.cleanup_failure is not None
    assert result.cleanup_failure["events"] == [
        {
            "operation": "scratch_chart_delete",
            "hresult": -2147000001,
            "message": "chart delete failed",
        }
    ]


def test_com_renderer_preserves_render_failure_when_global_cleanup_also_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class CopyPictureError(RuntimeError):
        hresult = -2146827284

    class QuitError(RuntimeError):
        hresult = -2147000002

    source = tmp_path / "source.xlsx"
    source.write_bytes(b"immutable source workbook")
    _install_fake_excel(
        monkeypatch,
        _FakeExcel(
            copy_error=CopyPictureError("CopyPicture failed"),
            quit_error=QuitError("Excel Quit failed"),
        ),
    )
    monkeypatch.setattr(render_runner.time, "sleep", lambda _seconds: None)
    targets = {"ANF": (RenderTarget("Protected", "A1:C5", "test"),)}

    result = _render_ranges_with_excel_com({"ANF": source}, tmp_path / "images", targets)[0]

    cleanup = next(row for row in result.operations if row["operation"] == "cleanup")
    assert result.status == "fail"
    assert result.failing_operation == "copy_picture"
    assert "CopyPicture failed" in result.message
    assert cleanup["status"] == "fail"
    assert result.cleanup_failure is not None
    assert [event["operation"] for event in result.cleanup_failure["events"]] == ["excel_quit"]
    assert result.cleanup_failure["hresult"] == -2147000002


def test_com_renderer_reports_excel_quit_failure_as_cleanup(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class QuitError(RuntimeError):
        hresult = -2147000003

    source = tmp_path / "source.xlsx"
    source.write_bytes(b"immutable source workbook")
    _install_fake_excel(monkeypatch, _FakeExcel(quit_error=QuitError("Excel Quit failed")))
    targets = {"ANF": (RenderTarget("Protected", "A1:C5", "test"),)}

    result = _render_ranges_with_excel_com({"ANF": source}, tmp_path / "images", targets)[0]

    assert result.status == "fail"
    assert result.failing_operation == "cleanup"
    assert result.owned_excel_process_cleanup == "FAIL"
    assert result.cleanup_failure is not None
    assert result.cleanup_failure["events"][0]["operation"] == "excel_quit"
    assert result.cleanup_failure["hresult"] == -2147000003


def test_com_renderer_reports_owned_image_deletion_failure_without_hiding_render_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"immutable source workbook")
    _install_fake_excel(monkeypatch, _FakeExcel())
    monkeypatch.setattr(render_runner, "_image_nonblank", lambda _path: (False, "blank image"))
    monkeypatch.setattr(
        render_runner,
        "_delete_owned_render_image",
        lambda _path: (_ for _ in ()).throw(PermissionError("owned image delete failed")),
    )
    targets = {"ANF": (RenderTarget("Protected", "A1:C5", "test"),)}

    result = _render_ranges_with_excel_com({"ANF": source}, tmp_path / "images", targets)[0]

    assert result.status == "fail"
    assert result.failing_operation == "nonblank_check"
    assert result.message == "blank image"
    assert result.cleanup_failure is not None
    assert result.cleanup_failure["events"][0]["operation"] == "temporary_file_delete"
    assert "owned image delete failed" in result.cleanup_failure["events"][0]["message"]
    assert result.image_path is not None
    result.image_path.unlink()


@pytest.mark.skipif(sys.platform != "win32", reason="Excel COM image rendering is Windows-only")
def test_real_excel_renders_protected_source_via_owned_scratch_workbook(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Protected"
    for row in range(1, 7):
        for column in range(1, 5):
            ws.cell(row, column, f"R{row}C{column}")
    ws.protection.sheet = True
    source = tmp_path / "protected.xlsx"
    wb.save(source)
    wb.close()
    source_bytes = source.read_bytes()
    targets = {"TEST": (RenderTarget("Protected", "A1:D6", "test"),)}

    result = _render_ranges_with_excel_com({"TEST": source}, tmp_path / "images", targets)[0]

    assert result.status == "pass", result.to_dict()
    assert result.image_path is not None and result.image_path.stat().st_size > 200
    assert source.read_bytes() == source_bytes
    assert result.source_bytes_unchanged is True
    assert result.owned_excel_process_cleanup == "PASS"
    assert result.cleanup_failure is None
    assert result.failing_operation == ""
    assert all(operation["status"] == "pass" for operation in result.operations)
    assert not list(tmp_path.glob("*.tmp"))


def test_openpyxl_style_validation_passes_current_legacy_core_surfaces() -> None:
    for ticker in TICKERS:
        report = validate_openpyxl_layout(
            _workbook_path(ticker),
            ticker,
            render_targets=_legacy_core_render_targets(ticker),
        )
        blocking = [issue for issue in report.issues if issue.severity == "error"]
        assert not blocking, f"{ticker}: openpyxl layout/style issues: {[issue.to_dict() for issue in blocking[:10]]}"
        assert report.checked_sheets >= 4
        assert report.max_row_height <= 95


def test_render_validation_skips_com_cleanly_and_still_runs_style_checks(tmp_path: Path) -> None:
    # PBI has populated manifest-era conditional sheets; GPRE/ANF canonical files
    # retain legacy debt layouts and are covered above through their core surfaces.
    workbooks = {"PBI": _workbook_path("PBI")}

    report = run_render_validation(
        workbooks,
        output_root=tmp_path,
        timestamp="unit",
        enable_com=False,
    )

    assert report.output_dir == tmp_path / "final_validation_unit"
    assert report.render_status == "skipped"
    assert "disabled" in report.skip_reason.lower()
    assert report.output_dir.exists()
    assert set(report.style_reports) == {"PBI"}
    assert all(not [issue for issue in style.issues if issue.severity == "error"] for style in report.style_reports.values())
    assert report.to_summary_rows()[0]["Overall"] in {"PASS", "SKIP_RENDER"}


def test_openpyxl_layout_validator_detects_broken_user_facing_styles(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation"
    ws["A1"] = "Valuation"
    ws["A2"] = "Metric"
    ws["B2"] = "Value"
    ws["A3"] = "Unstyled body"
    for sheet_name in ["PBI_Investment_Case", "Promise_Progress_UI", "Quarter_Notes_UI", "Operating_Drivers", "Needs_Review"]:
        sheet = wb.create_sheet(sheet_name)
        sheet["A1"] = sheet_name
    path = tmp_path / "PBI_model.xlsx"
    wb.save(path)

    report = validate_openpyxl_layout(path, "PBI")

    assert any(issue.severity == "error" for issue in report.issues)
    assert any("missing title/header styling" in issue.message.lower() for issue in report.issues)
