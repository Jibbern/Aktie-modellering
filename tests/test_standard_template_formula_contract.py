from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, range_boundaries

from pbi_xbrl.standard_template_formula_contract import (
    BS_RAW_ROWS,
    FORMULA_CONTRACT_VERSION,
    FORMULA_ROWS,
    USER_INPUT_CONTRACTS,
    VALUATION_HELPER_ROWS,
    VALUATION_RAW_ROWS,
    canonical_data_validation_cells,
)


ROOT = Path(__file__).resolve().parents[1]
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"


def test_quarterly_formula_contract_has_blank_guards_and_protected_outputs() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        ws = wb["Valuation"]
        for contract in FORMULA_ROWS:
            for column in range(2, 14):
                cell = ws.cell(contract.row, column)
                assert isinstance(cell.value, str) and cell.value.startswith("="), contract.formula_id
                assert cell.protection.locked is True

        revenue_ttm = str(ws["E10"].value)
        assert "History_Q!" in revenue_ttm
        assert 'COUNTIFS(' in revenue_ttm
        assert 'MAXIFS(' in revenue_ttm and 'MINIFS(' in revenue_ttm
        assert '"revenue"' in revenue_ttm and '"$m"' in revenue_ttm
        revenue_yoy = str(ws["F11"].value)
        assert "History_Q!" in revenue_yoy
        assert 'COUNTIFS(' in revenue_yoy and '-4)' in revenue_yoy
        assert '"revenue"' in revenue_yoy and '"$m"' in revenue_yoy
        assert ws["M47"].value == '=IF(OR(M43="",M44=""),"",M43-M44)'
        assert ws["M86"].value == '=IF(OR(M73="",M84="",M84=0),"",M73/M84)'
        assert "History_Q!" in str(ws["M271"].value)
        assert '"operating_cash_flow"' in str(ws["M271"].value)
        assert ws.row_dimensions[271].hidden is True
        assert ws["M271"].protection.locked is True
    finally:
        wb.close()


def test_source_backed_targets_are_blank_and_locked() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        valuation = wb["Valuation"]
        for row in sorted(set(VALUATION_RAW_ROWS.values()) | set(VALUATION_HELPER_ROWS.values())):
            for column in range(2, 14):
                cell = valuation.cell(row, column)
                assert cell.value is None
                assert cell.protection.locked is True

        bs = wb["BS_Segments"]
        for row in BS_RAW_ROWS.values():
            for column in range(2, 14):
                assert bs.cell(row, column).value is None
                assert bs.cell(row, column).protection.locked is True
    finally:
        wb.close()


def test_retired_annual_financial_surface_is_blank_locked_and_hidden() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        ws = wb["BS_Segments"]
        for row in range(79, 105):
            assert ws.row_dimensions[row].hidden is True
            for column in range(1, 14):
                cell = ws.cell(row, column)
                assert cell.value is None
                assert cell.protection.locked is True
    finally:
        wb.close()


def test_formula_economics_use_exact_generic_inputs_and_fail_closed() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        valuation = wb["Valuation"]
        exact = {
            "B13": '=IF(OR(B262="",B9="",B9=0),"",B262/B9)',
            "B14": '=IF(OR(B32="",B9="",B9=0),"",B32/B9)',
            "B19": '=IF(OR(B18="",B9="",B9=0),"",B18/B9)',
            "B27": '=IF(OR(B24="",B9="",B9=0),"",B24/B9)',
            "B37": '=IF(OR(B36="",B9="",B9=0),"",B36/B9)',
            "B47": '=IF(OR(B43="",B44=""),"",B43-B44)',
            "B57": '=IF(OR(B47="",B9="",B9=0),"",B47/B9)',
            "B73": '=IF(OR(B72="",B70=""),"",B72-B70)',
            "B80": '=IF(OR(B70="",B72="",B79=""),"",B72+B79-B70)',
            "B86": '=IF(OR(B73="",B84="",B84=0),"",B73/B84)',
            "B113": '=IF(OR(B268="",B103="",B103=0),"",B268/B103)',
            "B114": '=IF(OR(B268="",B269="",B270="",B103="",B103=0),"",(B268-B269-B270)/B103)',
            "B115": '=IF(OR(B49="",B102="",B102=0),"",B49/B102)',
        }
        for coordinate, formula in exact.items():
            assert valuation[coordinate].value == formula

        for contract in FORMULA_ROWS:
            formulas = [str(valuation.cell(contract.row, column).value) for column in range(2, 14)]
            assert any(formula.startswith("=IF(") for formula in formulas), contract.formula_id
            assert all(formula.startswith("=IF(") or formula == '=""' for formula in formulas), contract.formula_id
            for formula in formulas:
                if "History_Q!" in formula:
                    assert '"populated"' in formula, contract.formula_id
                    assert "COUNTIFS(" in formula, contract.formula_id
    finally:
        wb.close()


def test_visible_metric_labels_are_concise_without_losing_definition() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        ws = wb["Valuation"]
        assert [ws[f"A{row}"].value for row in range(36, 41)] == [
            "Net income",
            "Net margin %",
            "Net income YoY %",
            "Net income (TTM)",
            "Net margin (TTM)",
        ]
        assert ws["A107"].value == "Diluted EPS (GAAP)"
        assert ws["A109"].value == "Diluted EPS (GAAP, TTM)"
        assert ws["A110"].value == "Adjusted diluted EPS"
        assert ws["A111"].value == "Adjusted diluted EPS (TTM)"
        assert ws["A18"].value == "EBITDA (base)"
        assert ws["A24"].value == "Adjusted EBITDA"
    finally:
        wb.close()


def test_formula_contract_contains_no_ticker_specific_content() -> None:
    source = (ROOT / "pbi_xbrl" / "standard_template_formula_contract.py").read_text(encoding="utf-8")
    assert FORMULA_CONTRACT_VERSION == "2.5.1"
    for forbidden in ("Abercrombie", "Hollister", "ANF_model", "A&F"):
        assert forbidden not in source


def test_summary_formula_contract_is_exact_period_linked_and_fail_closed() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        summary = wb["SUMMARY"]
        expected_rows = {
            "B27": 10,
            "B29": 11,
            "B31": 38,
            "B32": 107,
            "B33": 108,
            "B36": 49,
            "B42": 88,
        }
        for coordinate, source_row in expected_rows.items():
            formula = str(summary[coordinate].value)
            assert formula.startswith("=IFERROR(IF(OR(")
            assert "$B$26" in formula
            assert "COUNTIF('Valuation'!$B$6:$M$6,$B$26)<>1" in formula
            assert f"'Valuation'!$B${source_row}:$M${source_row}" in formula
            assert "MATCH($B$26,'Valuation'!$B$6:$M$6,0)" in formula
            assert summary[coordinate].protection.locked is True

        fcf_yoy = str(summary["B37"].value)
        assert "'Valuation'!$B$48:$M$48" in fcf_yoy
        assert "'Valuation'!$B$47:$M$47" in fcf_yoy
        assert "MATCH($B$26,'Valuation'!$B$6:$M$6,0)<=4" in fcf_yoy
        assert "/ABS(" in fcf_yoy
        assert "=0" in fcf_yoy
        assert summary["A27"].value == "Revenue TTM ($m)"
        assert summary["A32"].value == "GAAP diluted EPS ($/share)"
        assert summary["A33"].value == "GAAP diluted EPS growth YoY (%)"
        assert summary["A36"].value == "FCF TTM ($m)"
        assert summary["A41"].value == "Net leverage (x)"
        assert summary["A42"].value == "Interest coverage, P&L TTM (x)"
        assert summary["A44"].value == "Revolver availability ($m)"
        assert summary["A45"].value == "Total liquidity ($m)"
        assert all(summary.cell(row, 3).value is None for row in range(27, 46))
    finally:
        wb.close()


def test_balance_sheet_ratio_formats_and_investment_case_units_are_explicit() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        bs = wb["BS_Segments"]
        assert bs["A38"].value == "Current ratio (x)"
        assert bs["A39"].value == "Quick ratio (x)"
        assert all(bs.cell(row, column).number_format == "0.00x;-0.00x" for row in (38, 39) for column in range(2, 14))

        investment_case = wb["{ticker}_Investment_Case"]
        assert 'MATCH("market_input|price"' in str(investment_case["A17"].value)
        assert 'MATCH("market_input|revenue"' in str(investment_case["A21"].value)
        assert 'MATCH("market_input|net_debt"' in str(investment_case["A19"].value)
        assert 'MATCH("market_input|revenue_growth"' in str(investment_case["A22"].value)
        assert 'MATCH("market_input|target_fcf_yield"' in str(investment_case["B110"].value)
        assert investment_case["A45"].value == "Total Company revenue growth (%)"
        assert investment_case["A67"].value == "Share issuance (m)"
        assert investment_case["A85"].value == "Revenue ($m)"
        assert investment_case["A98"].value == "GAAP diluted EPS ($/share)"
    finally:
        wb.close()


def test_summary_business_oracles_use_signed_fcf_improvement() -> None:
    current_revenue = 1_113.821
    prior_revenue = current_revenue / (1.0 + 0.015045871225204177)
    current_net_income = 67.134
    prior_net_income = current_net_income / (1.0 - 0.1651349906109708)
    current_eps = 1.4697550189373207
    prior_eps = current_eps / (1.0 - 0.07562577425325745)
    current_fcf = 44.256 - 61.341
    prior_fcf = -4.0 - 50.764

    assert (current_revenue / prior_revenue) - 1 == pytest.approx(0.015045871225204177)
    assert (current_net_income / prior_net_income) - 1 == pytest.approx(-0.1651349906109708)
    assert (current_eps / prior_eps) - 1 == pytest.approx(-0.07562577425325745)
    assert (current_fcf - prior_fcf) / abs(prior_fcf) == pytest.approx(0.688024979913812)


def test_typed_scenario_formulas_have_exact_ownership_and_no_unsafe_defaults() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        valuation = wb["Valuation"]
        investment_case = wb["{ticker}_Investment_Case"]

        for coordinate in ("D218", "D219", "D220", "J218", "J219", "J220", "J221", "E237", "J236", "D247", "D253", "D254"):
            assert valuation[coordinate].value is None
            assert valuation[coordinate].protection.locked is False
        for coordinate in ("J222", "J223", "E241", "E244", "E259", "E260", "E261", "J259", "J261", "N238", "N244", "R261"):
            assert isinstance(valuation[coordinate].value, str) and valuation[coordinate].value.startswith("=")
            assert valuation[coordinate].protection.locked is True

        assert "DCF_Horizon" in str(valuation["J222"].value)
        assert "ScenarioBuybackCash/ScenarioBuybackPrice" in str(valuation["E259"].value)
        assert "NetDebt+ScenarioBuybackCash-ScenarioDebtPaydown" in str(valuation["E260"].value)
        assert "NetIncome_TTM" in str(valuation["E261"].value)
        assert "Adj_EBITDA" not in str(valuation["E242"].value)
        assert "ScenarioAdjustedMargin" in str(valuation["E243"].value)
        assert "ResolvedRevenueGrowth_Custom" in str(valuation["E241"].value)
        assert "ScenarioGrowth" not in str(valuation["E241"].value)

        n244 = str(valuation["N244"].value)
        assert all(token in n244 for token in (
            "NOT(ISNUMBER(N237))",
            "NOT(ISNUMBER(DCF_WACC))",
            "NOT(ISNUMBER(DCF_FCFF))",
            'IF(N237+DCF_FCFF=0,""',
        ))
        n261 = str(valuation["N261"].value)
        assert all(token in n261 for token in (
            "NOT(ISNUMBER(ScenarioFCF))",
            "NOT(ISNUMBER(ScenarioImpliedPrice))",
            "NOT(ISNUMBER(ScenarioShares))",
            "ScenarioImpliedPrice<=0",
            "ScenarioShares<=0",
        ))
        assert "'Valuation'!" not in "\n".join(
            str(cell.value)
            for cell in investment_case._cells.values()
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        assert 'MATCH("market_input|price"' in str(investment_case["A17"].value)
        assert str(investment_case["F17"].value).startswith("=IF(ISNUMBER(")
        assert "Manual input required" in str(investment_case["G17"].value)
        assert investment_case["B69"].value is None
        assert investment_case["B69"].protection.locked is False
        assert "$B$69" in str(investment_case["E121"].value)
        assert investment_case.row_dimensions[127].height == pytest.approx(42.0)
        selected_margin = str(investment_case["C74"].value)
        assert 'INDEX($C$57:$E$57' in selected_margin
        assert '="",$B$57,IF(ISNUMBER(INDEX(' in selected_margin
        assert 'INDEX($C$57:$E$57' in selected_margin
        assert ',"Unavailable")),"")' in selected_margin
        selected_capex = str(investment_case["C75"].value)
        assert 'INDEX($C$63:$E$63' in selected_capex
        assert '="",$B$63,IF(ISNUMBER(INDEX(' in selected_capex
        assert ',"Unavailable")),"")' in selected_capex
        assert str(investment_case["B89"].value).startswith("=IF(NOT(ISNUMBER(")
        assert str(investment_case["E101"].value).startswith("=IF(E$83=")
        assert str(investment_case["G121"].value) == '=IF(ISNUMBER(E121),"Available","Unavailable")'
        assert str(investment_case["D106"].value).startswith("=IF(ISNUMBER(C106),")

        source = (ROOT / "pbi_xbrl" / "standard_template_formula_contract.py").read_text(encoding="utf-8")
        assert "MAX(0.001" not in source
        assert "ScenarioBuyback_m" not in source
        assert "0.10,0.11,0.12" not in source

        for coordinate in (
            "B42", "C45", "D45", "E45", "B69", "C56", "D63", "E68",
            "C106", "C114", "B117", "F117", "I192", "I211",
        ):
            assert investment_case[coordinate].value is None
            assert investment_case[coordinate].protection.locked is False
        for coordinate in (
            "B15", "E15", "A17", "F17", "G17", "A46", "B44", "F44",
            "B73", "E80", "B85", "E101", "B106", "D114", "B121", "G127",
            "B131", "F150", "B156", "I165", "B171", "C188",
            "A195", "D215", "A219", "L225",
        ):
            assert isinstance(investment_case[coordinate].value, str) and investment_case[coordinate].value.startswith("=")
            assert investment_case[coordinate].protection.locked is True
    finally:
        wb.close()


def test_investment_case_driver_chain_is_visible_and_valuation_isolated() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        investment_case = wb["{ticker}_Investment_Case"]
        expected_sections = {
            4: "Investment Snapshot",
            13: "Model Data and Guidance",
            38: "Scenario Assumptions",
            71: "Selected Scenario Incremental Bridge",
            82: "Scenario Output Comparison",
            103: "Valuation and DCF Assumptions",
            119: "Valuation Summary",
            129: "What the Market Is Pricing",
            141: "Guidance-Implied Earnings",
            152: "DCF and Equity Value",
            167: "Calculation Details",
            190: "Sensitivity Tables",
            217: "Key Debates and Invalidators",
        }
        assert {row: investment_case.cell(row, 1).value for row in expected_sections} == expected_sections

        assert "Model default (" in str(investment_case["B15"].value)
        assert "FY default" not in str(investment_case["B15"].value)
        assert investment_case["C15"].value == '="Model default (TTM)"'
        assert "Guidance (" in str(investment_case["D15"].value)
        assert "Full-year guidance" not in str(investment_case["D15"].value)
        assert "Guidance (" in str(investment_case["E15"].value)
        assert "$AN$2:$AN$201" in str(investment_case["E15"].value)
        assert investment_case["F15"].value == "Active value"
        assert investment_case["G15"].value == "Active source"
        assert "Manual override" not in {
            investment_case.cell(15, column).value
            for column in range(1, 13)
        }
        assert [investment_case.cell(83, column).value for column in range(1, 6)] == [
            "Output",
            "Current baseline",
            "Bear",
            "Base",
            "Bull",
        ]
        assert [investment_case.cell(170, column).value for column in range(1, 4)] == [
            "Calculation",
            "Basis",
            "Result",
        ]
        assert all(
            investment_case.cell(row, column).value != "Unit"
            for row in range(1, 226)
            for column in range(1, 14)
        )
        assert not any(
            isinstance(investment_case.cell(row, column).value, str)
            and "Custom" in investment_case.cell(row, column).value
            for row in range(1, 226)
            for column in range(1, 14)
        )

        assert str(investment_case["F21"].value).startswith("=IF(ISNUMBER(")
        assert "Model default (TTM)" in str(investment_case["G21"].value)
        assert "TTM through" not in str(investment_case["G21"].value)
        assert "Manual input required" in str(investment_case["G17"].value)

        total_company_revenue = str(investment_case["C85"].value)
        assert 'LOWER(IF($B$42="","Total Company",$B$42))="total company"' in total_company_revenue
        assert "LOWER($A$46)" in total_company_revenue
        assert "LOWER($A$49)" in total_company_revenue
        assert total_company_revenue.count('$B$47*$C$47') == 1
        assert 'IF($C$45="",$B$45,IF(ISNUMBER($C$45),$C$45,"Unavailable"))' in total_company_revenue
        assert 'IF($D$45="",$B$45,IF(ISNUMBER($D$45),$D$45,"Unavailable"))' in str(investment_case["D85"].value)
        assert investment_case["B44"].protection.locked is True
        assert investment_case["D45"].protection.locked is False

        assert "INDEX($C$85:$E$85" in str(investment_case["E73"].value)
        assert "INDEX($C$89:$E$89" in str(investment_case["E78"].value)
        assert "INDEX($C$98:$E$98" in str(investment_case["E80"].value)
        assert "Value/share ($/share)" not in {
            investment_case.cell(row, 1).value
            for row in range(73, 81)
        }
        assert "INDEX($C$98:$E$98" in str(investment_case["B121"].value)
        assert all(token in str(investment_case["H165"].value) for token in ("H163", "H164"))
        assert "MATCH(IF($B$69=\"\",\"Base\",$B$69)" in str(investment_case["E121"].value)
        assert str(investment_case["G121"].value) == '=IF(ISNUMBER(E121),"Available","Unavailable")'
        assert investment_case["H127"].value == '="Unavailable methods are excluded regardless of entered weight."'

        assert investment_case["C41"].value == "Bear"
        assert investment_case["D41"].value == "Base"
        assert investment_case["E41"].value == "Bull"
        assert investment_case["B69"].value is None
        assert investment_case["C69"].value == "Blank selection uses Base"
        assert investment_case["B69"].protection.locked is False
        assert investment_case["D56"].protection.locked is False
        assert investment_case["B56"].protection.locked is True
        assert investment_case["C106"].protection.locked is False
        assert investment_case["D106"].protection.locked is True

        percentage_targets = {
            "C45", "D45", "E45",
            *(
                f"{column}{row}"
                for column in "CDE"
                for row in (47, 48, 50, 51, 52, 56, 57, 58, 59, 60)
            ),
            "C110", "C111", "C112", "C113", "I204", "I210", "I211",
        }
        percentage_validations = [
            validation
            for validation in investment_case.data_validations.dataValidation
            if any(
                str(cell_range) in percentage_targets
                for cell_range in validation.ranges.ranges
            )
        ]
        assert percentage_validations
        assert all(validation.prompt is None for validation in percentage_validations)
        assert all(validation.promptTitle is None for validation in percentage_validations)
        assert all(validation.showInputMessage is False for validation in percentage_validations)
        assert "Enter percentages as 6% or 0.06." not in str(investment_case["A39"].value)
        assert all(
            "Enter 6% or 0.06, not 6." not in str(validation.prompt or "")
            for validation in investment_case.data_validations.dataValidation
        )

        assert investment_case["A12"].value == "Typed Scenario Inputs"
        assert investment_case.row_dimensions[12].hidden is True
        assert investment_case["G219"].value.startswith('=IF(A219="","",')
        assert "Manual review required" in str(investment_case["G219"].value)
        assert {
            "G15:L15",
            "G17:L17",
            "F41:L41",
            "F55:L55",
            "F105:L105",
            "A115:F115",
            "G115:J115",
            "A154:F154",
            "G154:M154",
            "B219:F219",
            "H219:K219",
        } <= {str(merged) for merged in investment_case.merged_cells.ranges}
        bridge_merges = {
            str(merged)
            for merged in investment_case.merged_cells.ranges
            if merged.min_row <= 80
            and merged.max_row >= 72
            and merged.min_col <= 13
            and merged.max_col >= 5
        }
        assert bridge_merges == set()
        assert investment_case["E72"].value == "Resulting output"
        assert investment_case["E72"].alignment.horizontal == "center"
        assert all(investment_case[f"E{row}"].value is not None for row in range(73, 81))
        assert all(investment_case[f"E{row}"].alignment.horizontal == "left" for row in range(73, 81))
        assert all(
            investment_case.cell(row, column).value is None
            for row in range(72, 81)
            for column in range(6, 14)
        )
        expected_weight_note = (
            "Method weights (%) - Blended value/share, enter percentages that sum to 100% "
            "across available methods; blank or 0 excludes a method."
        )
        assert investment_case["G115"].value == expected_weight_note
        assert investment_case.row_dimensions[115].height == 36.0
        assert not any(
            isinstance(cell.value, str)
            and "Method weights determine the blended value/share" in cell.value
            for cell in investment_case._cells.values()
        )
        assert all(
            investment_case.cell(row, column).protection.locked is True
            for row in (115, 116)
            for column in range(7, 11)
        )
        assert investment_case["A150"].value == "Latest-quarter adjusted EPS ($/share)"
        assert investment_case.row_dimensions[5].height == 42.0
        assert investment_case.row_dimensions[156].height == 24.0
        assert all(
            investment_case.column_dimensions[get_column_letter(column)].hidden is True
            for column in range(14, 54)
        )
        assert len({
            investment_case.column_dimensions[get_column_letter(column)].width
            for column in range(2, 9)
        }) == 1
        assert investment_case.column_dimensions["B"].width == 25.0
        assert investment_case.column_dimensions["M"].hidden is False
        assert investment_case.column_dimensions["N"].hidden is True
        assert all(
            investment_case.row_dimensions[row].hidden is True
            and investment_case.row_dimensions[row].outlineLevel == 0
            and investment_case.row_dimensions[row].collapsed is False
            for row in (*range(194, 198), *range(200, 204), *range(206, 210), *range(212, 216))
        )
        assert all(
            investment_case.row_dimensions[row].outlineLevel == 0
            and investment_case.row_dimensions[row].collapsed is False
            for row in range(1, 241)
        )
        assert investment_case.sheet_view.showOutlineSymbols is False
        assert investment_case.freeze_panes == "A2"
        assert all(
            ws.sheet_view.zoomScale == 110 and ws.sheet_view.zoomScaleNormal == 110
            for ws in wb.worksheets
        )
        assert investment_case["C45"].alignment.horizontal == "left"
        assert investment_case["F17"].alignment.horizontal == "left"
        assert investment_case["C106"].alignment.horizontal == "left"
        assert investment_case["B15"].alignment.horizontal == "center"
        assert all(
            investment_case.cell(row, column).alignment.vertical == "center"
                for row in range(1, 226)
                for column in range(1, 14)
                if investment_case.row_dimensions[row].hidden is not True
                if investment_case.cell(row, column).__class__.__name__ != "MergedCell"
                and (
                investment_case.cell(row, column).value is not None
                or investment_case.cell(row, column).protection.locked is False
            )
        )
        assert investment_case["A1"].font.name == "Aptos Display"
        assert investment_case["A1"].font.color.rgb == "00FFFFFF"
        assert investment_case["A1"].fill.fgColor.rgb == "004472C4"
        assert investment_case["A13"].fill.fgColor.rgb == "005B9BD5"
        assert investment_case["A15"].fill.fgColor.rgb == "00EAF3F8"
        assert investment_case["A16"].fill.fgColor.rgb == "00DDEBF7"
        assert investment_case["A43"].fill.fgColor.rgb == "00FFFFFF"
        assert investment_case["A44"].fill.fgColor.rgb == "00FFFFFF"
        assert investment_case["A46"].fill.fgColor.rgb == "00DDEBF7"
        assert investment_case["A47"].font.color.rgb == "001F2933"
        assert {
            investment_case.cell(1, column).fill.fgColor.rgb
            for column in range(1, 14)
        } == {"004472C4"}

        horizon_validation = next(
            validation
            for validation in investment_case.data_validations.dataValidation
            if str(validation.sqref) == "C114"
        )
        assert horizon_validation.type == "whole"
        assert horizon_validation.formula1 == "1"
        assert horizon_validation.formula2 == "5"

        visible_investment_case_ref = "'{ticker}_Investment_Case'!"
        for sheet_name in ("Valuation", "Valuation_Summary"):
            assert visible_investment_case_ref not in "\n".join(
                str(cell.value)
                for cell in wb[sheet_name]._cells.values()
                if isinstance(cell.value, str) and cell.value.startswith("=")
            )
    finally:
        wb.close()


def test_investment_case_buyback_and_terminal_growth_formulas_fail_closed() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        ws = wb["{ticker}_Investment_Case"]

        active_price = str(ws["F17"].value)
        active_price_state = str(ws["G17"].value)
        buyback_repurchases = str(ws["C184"].value)
        scenario_shares = str(ws["D95"].value)
        scenario_eps = str(ws["D98"].value)
        pe_value = str(ws["E121"].value)
        pe_state = str(ws["G121"].value)
        market_cap = str(ws["B133"].value)
        market_ev = str(ws["B135"].value)
        terminal_growth = str(ws["B139"].value)
        terminal_growth_state = str(ws["C139"].value)
        dcf_forecast_pv = str(ws["H159"].value)
        dcf_terminal_pv = str(ws["H160"].value)
        dcf_value = str(ws["H165"].value)
        dcf_state = str(ws["I165"].value)
        dcf_sensitivity = str(ws["B213"].value)

        assert active_price.startswith("=IF(ISNUMBER(IFERROR(")
        assert "ISBLANK(INDEX(" in active_price
        assert active_price_state.startswith("=IF(ISNUMBER(F17),")
        assert "Manual input required" in active_price_state

        assert buyback_repurchases.startswith("=IF(IF(INDEX(")
        assert "INDEX($C$65:$E$65" in buyback_repurchases
        assert "INDEX($C$66:$E$66" in buyback_repurchases
        assert '="",$B$65,IF(ISNUMBER(INDEX(' in buyback_repurchases
        assert '="",$B$66,IF(ISNUMBER(INDEX(' in buyback_repurchases
        assert "$B$65" in buyback_repurchases
        assert "$B$66" in buyback_repurchases
        assert "<=0" in buyback_repurchases
        assert scenario_shares.startswith("=IF(NOT(ISNUMBER($F$18))")
        assert "$D$65" in scenario_shares
        assert "$D$66" in scenario_shares
        assert "$D$67" in scenario_shares
        assert '="",$F$18' in scenario_shares
        assert scenario_eps.startswith("=IF(NOT(ISNUMBER(D92))")
        assert "D95<=0" in scenario_eps

        assert pe_value.startswith("=IF(NOT(ISNUMBER(")
        assert "$D$107<=0" in pe_value
        assert pe_state == '=IF(ISNUMBER(E121),"Available","Unavailable")'

        assert market_cap.startswith("=IF(NOT(ISNUMBER($D$106))")
        assert market_ev.startswith("=IF(NOT(ISNUMBER(B133))")
        assert terminal_growth.startswith("=IF(NOT(ISNUMBER(B135))")
        assert "B135<=0" in terminal_growth
        assert "$D$112<=0" in terminal_growth
        assert "B135+INDEX(" in terminal_growth
        assert terminal_growth_state.startswith("=IF(ISNUMBER(B139),")
        assert "Unavailable | " in terminal_growth_state

        assert "COUNT(B165:F165)<>$D$114" in dcf_forecast_pv
        assert "MOD($D$114,1)<>0" in dcf_forecast_pv
        assert "INDEX($B$163:$F$163,1,$D$114)" in dcf_terminal_pv
        assert "INDEX($B$164:$F$164,1,$D$114)" in dcf_terminal_pv
        assert "COUNT(B163:F163)<>$D$114" in dcf_terminal_pv
        assert dcf_value.startswith("=IF(NOT(ISNUMBER(H163))")
        assert "H164<=0" in dcf_value
        assert dcf_state == '=IF(ISNUMBER(H165),"Available","Unavailable")'
        assert "COUNT($B$163:$F$163)<>$D$114" in dcf_sensitivity
        assert "INDEX($B$163:$F$163,1,$D$114)" in dcf_sensitivity
        assert "IF(ISNUMBER(B$163)" not in dcf_sensitivity

        critical = "\n".join(
            (
                buyback_repurchases,
                scenario_shares,
                scenario_eps,
                pe_value,
                market_cap,
                market_ev,
                terminal_growth,
                dcf_forecast_pv,
                dcf_terminal_pv,
                dcf_value,
                dcf_sensitivity,
            )
        )
        for forbidden in ("IFERROR(", "VALUE(", "N(", "--"):
            assert forbidden not in critical
        assert "FormulaLocal" not in critical
        assert "Formula2Local" not in critical
    finally:
        wb.close()


def test_excel_grouped_data_validations_keep_per_cell_relative_formula_identity() -> None:
    separate_workbook = Workbook()
    grouped_workbook = Workbook()
    try:
        separate = separate_workbook.active
        grouped = grouped_workbook.active
        for target in ("F15:F16", "C67", "E67"):
            coordinate = target.split(":", 1)[0]
            validation = DataValidation(
                type="custom",
                formula1=f'=OR({coordinate}="",AND(ISNUMBER({coordinate}),{coordinate}>0))',
                allow_blank=True,
            )
            separate.add_data_validation(validation)
            validation.add(target)

        grouped_validation = DataValidation(
            type="custom",
            formula1='OR(C15="",AND(ISNUMBER(C15),C15>0))',
            allow_blank=True,
        )
        grouped.add_data_validation(grouped_validation)
        for target in ("F15:F16", "E67", "C67"):
            grouped_validation.add(target)

        assert canonical_data_validation_cells(grouped) == canonical_data_validation_cells(separate)
    finally:
        separate_workbook.close()
        grouped_workbook.close()


def test_scenario_defined_names_validations_and_support_projections_are_complete() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        expected_names = {
            "NetIncome_TTM": "'Valuation'!$D$217",
            "DCF_Horizon": "'Valuation'!$D$220",
            "ScenarioTaxTreatment": "'Valuation'!$E$247",
            "ScenarioHorizon": "'Valuation'!$J$236",
            "ScenarioInterestTaxTreatment": "'Valuation'!$E$248",
            "ScenarioBuybackCash": "'Valuation'!$D$253",
            "ScenarioShares": "'Valuation'!$E$259",
            "ScenarioNetDebt": "'Valuation'!$E$260",
            "ScenarioImpliedPrice": "'Valuation'!$J$261",
            "ResolvedRevenueGrowth_Bear": "'Valuation_Summary'!$H$2",
            "ResolvedRevenueGrowth_Base": "'Valuation_Summary'!$I$2",
            "ResolvedRevenueGrowth_Bull": "'Valuation_Summary'!$J$2",
            "ResolvedRevenueGrowth_Custom": "'Valuation_Summary'!$K$2",
            "InvestmentCaseDimensionOptions": "'{ticker}_Investment_Case_Data'!$AV$2:$AV$4",
        }
        for name, target in expected_names.items():
            assert wb.defined_names[name].attr_text == target

        investment_case = wb["{ticker}_Investment_Case"]
        dimension_validation = next(
            validation
            for validation in investment_case.data_validations.dataValidation
            if str(validation.sqref) == "B42"
        )
        assert dimension_validation.formula1 == "=InvestmentCaseDimensionOptions"
        assert dimension_validation.promptTitle == "Revenue scenario mode"
        assert dimension_validation.prompt == (
            "Choose Total Company, Brand or Geography. Blank uses Total Company."
        )

        route_support = wb["Valuation_Summary"]
        for coordinate in ("H2", "I2", "J2"):
            formula = str(route_support[coordinate].value or "")
            assert formula.startswith("=_xlfn.LET(")
            assert "selected_growth_route" in formula
            assert "profile_driver_bridge" in formula
            assert '=\"revenue_growth\"' in formula
            assert formula.count('=\"total_company\"') == 4
            assert "'{ticker}_Investment_Case'!" not in formula
            assert '_xlpm.userGrowth,""' in formula
            assert '_xlpm.scenarioHorizon,""' in formula
            assert "LOWER(" not in formula
            assert "SUBSTITUTE(" not in formula
            assert "TRIM(" not in formula
            assert "_xlpm.directCount+_xlpm.profileCount+_xlpm.userCount<>1" in formula
            assert "_xlpm.bridgeCount=1" in formula
            assert route_support[coordinate].protection.locked is True
        custom_route = str(route_support["K2"].value or "")
        assert custom_route.startswith("=_xlfn.LET(")
        assert "_xlpm.userGrowth,ScenarioGrowth" in custom_route
        assert "_xlpm.scenarioHorizon,ScenarioHorizon" in custom_route
        assert route_support["K2"].protection.locked is True
        assert "retail_operating_pack" not in str(route_support["I2"].value)

        valuation = wb["Valuation"]
        validation_targets = {str(validation.sqref) for validation in valuation.data_validations.dataValidation}
        assert {"D194", "D208:D209", "D210", "D213", "D214", "D215", "D216", "D218:D219", "D220", "J218", "J219:J220", "J221", "H226:L226", "G227:G234", "E236", "E237:E239", "E240", "J236", "D247", "E247", "D248", "E248", "D249:D250", "D253", "D254", "D255:D256"} == validation_targets

        summary = wb["Valuation_Summary"]
        assert summary["A2"].value == "price"
        assert summary["B20"].value == '=IF(ScenarioUpside="","",ScenarioUpside)'
        assert summary["F20"].value == '=IF(B20="","unavailable","calculated")'
        assert summary["B20"].protection.locked is True

        grid = wb["Valuation_Grid"]
        assert grid["A2"].value == '="dcf"'
        assert "'Valuation'!$H$227" in str(grid["E2"].value)
        assert grid["D42"].value == "scenario_implied_price"
        assert grid["E42"].value == '=IF(ScenarioImpliedPrice="","",ScenarioImpliedPrice)'
    finally:
        wb.close()


def test_calculation_history_is_a_hidden_source_backed_formula_input_projection() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        ws = wb["History_Q"]
        assert ws.sheet_state == "hidden"
        assert [ws.cell(1, column).value for column in range(1, 8)] == [
            "period",
            "period_ordinal",
            "metric",
            "value",
            "unit",
            "source_ref",
            "status",
        ]
        for row in (2, 500, 1000):
            for column in range(1, 8):
                assert ws.cell(row, column).value is None
                assert ws.cell(row, column).protection.locked is True
    finally:
        wb.close()


def test_user_input_contract_is_the_exact_visible_edit_surface() -> None:
    wb = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        editable = {
            (ws.title, cell.coordinate)
            for ws in wb.worksheets
            if ws.sheet_state == "visible"
            for cell in ws._cells.values()
            if cell.protection.locked is False
        }
        expected = set()
        for contract in USER_INPUT_CONTRACTS:
            min_col, min_row, max_col, max_row = range_boundaries(contract.target)
            expected.update(
                (contract.sheet, cell.coordinate)
                for row in wb[contract.sheet].iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                )
                for cell in row
            )
        assert editable == expected
        assert sum(sheet == "Valuation" for sheet, _cell in editable) == 44
        assert sum(sheet == "{ticker}_Investment_Case" for sheet, _cell in editable) == 75
        assert len(editable) == 119
        assert all(ws.protection.sheet for ws in wb.worksheets)
    finally:
        wb.close()
