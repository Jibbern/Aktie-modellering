from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import pytest

from pbi_xbrl.json_schema_validation import load_json_strict
from pbi_xbrl.standard_template_formula_contract import (
    USER_INPUT_CONTRACTS,
    validate_workbook_protection_contract,
)
from scripts.materialize_standard_template_shell import materialize_shell


ROOT = Path(__file__).resolve().parents[1]
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
SHELL_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"
MODULE_MANIFEST = ROOT / "docs" / "workbook_module_manifest.json"


def _editable_cells(workbook) -> set[tuple[str, str]]:
    return {
        (ws.title, cell.coordinate)
        for ws in workbook.worksheets
        for cell in ws._cells.values()
        if cell.protection.locked is False
    }


def _expected_full_union_inputs(workbook) -> set[tuple[str, str]]:
    result: set[tuple[str, str]] = set()
    for contract in USER_INPUT_CONTRACTS:
        min_col, min_row, max_col, max_row = range_boundaries(contract.target)
        result.update(
            (contract.sheet, cell.coordinate)
            for row in workbook[contract.sheet].iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            )
            for cell in row
        )
    return result


def test_checked_in_shell_has_exact_protected_editable_surface() -> None:
    manifest = load_json_strict(SHELL_MANIFEST)
    workbook = load_workbook(SHELL, data_only=False, read_only=False)
    try:
        enabled = manifest["module_profile"]["enabled_formula_ids"]
        assert validate_workbook_protection_contract(workbook, enabled) == []
        assert all(ws.protection.sheet for ws in workbook.worksheets)
        assert len([ws for ws in workbook.worksheets if ws.sheet_state == "visible"]) == 10
        assert len([ws for ws in workbook.worksheets if ws.sheet_state != "visible"]) == 34
        editable = _editable_cells(workbook)
        assert editable == _expected_full_union_inputs(workbook)
        assert len(editable) == 75
        assert sum(sheet == "Valuation" for sheet, _cell in editable) == 0
        assert sum(sheet == "{ticker}_Investment_Case" for sheet, _cell in editable) == 75
        assert workbook["History_Q"]["A2"].protection.locked is True
        assert workbook["Valuation"]["B9"].protection.locked is True
        assert workbook["BS_Segments"]["B9"].protection.locked is True
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("profile_id", "expected_editable_count"),
    (
        ("full_union", 75),
        ("anf", 75),
        ("pbi", 75),
        ("gpre", 75),
        ("core_only", 0),
    ),
)
def test_profile_materialization_resolves_protection_before_planning(
    tmp_path: Path,
    profile_id: str,
    expected_editable_count: int,
) -> None:
    output = tmp_path / f"{profile_id}.xlsx"
    manifest_output = tmp_path / f"{profile_id}.manifest.json"
    binding_output = tmp_path / f"{profile_id}.bindings.json"
    materialize_shell(
        data_root=tmp_path,
        output_path=output,
        manifest_path=SHELL_MANIFEST,
        binding_map_path=BINDING_MAP,
        module_manifest_path=MODULE_MANIFEST,
        module_profile_id=profile_id,
        contract_manifest_output_path=manifest_output,
        contract_binding_map_output_path=binding_output,
        update_identity=True,
    )
    projected_manifest = load_json_strict(manifest_output)
    workbook = load_workbook(output, data_only=False, read_only=False)
    try:
        enabled = projected_manifest["module_profile"]["enabled_formula_ids"]
        assert validate_workbook_protection_contract(workbook, enabled) == []
        assert all(ws.protection.sheet for ws in workbook.worksheets)
        assert len(_editable_cells(workbook)) == expected_editable_count
        if profile_id == "core_only":
            assert _editable_cells(workbook) == set()
    finally:
        workbook.close()
