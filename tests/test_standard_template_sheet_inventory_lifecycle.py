from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def _data_root() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = parent / "StockModelData"
        if candidate.exists():
            return candidate
    return ROOT.parent / "StockModelData"


DATA_ROOT = _data_root()
INVENTORY_JSON = ROOT / "docs" / "standard_template_sheet_inventory.json"
LIFECYCLE_JSON = ROOT / "docs" / "support_sheet_lifecycle_contract.json"
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
SOURCE_WORKBOOKS = {
    "PBI": DATA_ROOT / "outputs" / "Excel stock models" / "PBI_model.xlsx",
    "GPRE": DATA_ROOT / "outputs" / "Excel stock models" / "GPRE_model.xlsx",
    "ANF": DATA_ROOT / "outputs" / "Excel stock models" / "ANF_model.xlsx",
}
if not SOURCE_WORKBOOKS["GPRE"].exists():
    SOURCE_WORKBOOKS["GPRE"] = DATA_ROOT / "outputs" / "Excel stock models" / "GPRE_model.xlsm"
ALLOWED_CLASSIFICATIONS = {
    "standard_visible_shell_sheet",
    "required_support_shell_sheet",
    "optional_module_shell_sheet",
    "fixture_capacity_shell_sheet",
    "external_detail_sheet",
    "rejected_redundant_sheet",
    "legacy_module_source_sheet",
    "runtime_generated_support_sheet",
    "runtime_generated_audit_sheet",
    "optional_sector_pack_sheet",
    "ticker_specific_sheet",
    "deprecated_legacy_sheet",
    "exclude_from_standard_shell",
}
REQUIRED_FIELDS = {
    "sheet_name",
    "classification",
    "present_in_standard_shell",
    "present_in_PBI",
    "present_in_GPRE",
    "present_in_ANF",
    "standard_shell_state",
    "PBI_state",
    "GPRE_state",
    "ANF_state",
    "reason",
    "visible_formula_or_binding_dependency",
    "runtime_must_create_or_fill",
}
LIFECYCLE_FIELDS = {
    "sheet_name",
    "owner",
    "lifecycle",
    "neutral_shell_required",
    "headers_required",
    "allowed_writable_zones",
    "source_of_data",
    "created_when",
    "visibility",
    "validation_rules",
}


def _source_sheet_names() -> set[str]:
    names: set[str] = set()
    for path in SOURCE_WORKBOOKS.values():
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=True, data_only=False)
        try:
            names.update(wb.sheetnames)
        finally:
            wb.close()
    return names


def test_sheet_inventory_covers_every_pbi_gpre_anf_sheet() -> None:
    payload = json.loads(INVENTORY_JSON.read_text(encoding="utf-8"))
    rows = {row["sheet_name"]: row for row in payload["sheets"]}

    missing = _source_sheet_names() - set(rows)
    assert missing == set()

    for row in rows.values():
        assert REQUIRED_FIELDS <= set(row)
        assert row["classification"] in ALLOWED_CLASSIFICATIONS
        assert isinstance(row["reason"], str) and row["reason"].strip()


def test_support_lifecycle_contract_covers_non_visible_inventory_outputs() -> None:
    inventory = json.loads(INVENTORY_JSON.read_text(encoding="utf-8"))
    lifecycle = json.loads(LIFECYCLE_JSON.read_text(encoding="utf-8"))
    lifecycle_rows = {row["sheet_name"]: row for row in lifecycle["support_sheets"]}

    lifecycle_required_classes = {
        "required_support_shell_sheet",
        "runtime_generated_support_sheet",
        "runtime_generated_audit_sheet",
        "optional_sector_pack_sheet",
        "optional_module_shell_sheet",
        "fixture_capacity_shell_sheet",
        "external_detail_sheet",
    }
    for row in inventory["sheets"]:
        if row["classification"] not in lifecycle_required_classes:
            continue
        assert row["sheet_name"] in lifecycle_rows

    for row in lifecycle_rows.values():
        assert LIFECYCLE_FIELDS <= set(row)
        assert row["owner"] in {
            "frozen_shell",
            "value_only_runtime",
            "legacy_writer",
            "optional_sector_pack",
            "external_normalized_json",
        }
        assert row["lifecycle"] in {
            "static_template",
            "runtime_output",
            "audit_output",
            "source_cache_projection",
            "optional_sector_output",
            "external_detail",
        }


def test_required_support_shell_sheets_are_present_and_runtime_sheets_are_absent() -> None:
    inventory = json.loads(INVENTORY_JSON.read_text(encoding="utf-8"))
    lifecycle = json.loads(LIFECYCLE_JSON.read_text(encoding="utf-8"))
    required_support = {
        row["sheet_name"]
        for row in inventory["sheets"]
        if row["classification"] == "required_support_shell_sheet"
    }
    runtime_outputs = {
        row["sheet_name"]
        for row in lifecycle["support_sheets"]
        if row["owner"] == "value_only_runtime" and not row["neutral_shell_required"]
    }

    wb = load_workbook(SHELL, read_only=True, data_only=False)
    try:
        shell_sheets = set(wb.sheetnames)
    finally:
        wb.close()

    assert required_support <= shell_sheets
    assert runtime_outputs.isdisjoint(shell_sheets - required_support)


def test_gpre_sector_pack_sheets_are_optional_not_standard() -> None:
    payload = json.loads(INVENTORY_JSON.read_text(encoding="utf-8"))
    rows = {row["sheet_name"]: row for row in payload["sheets"]}

    for sheet_name in ("Economics_Overlay", "Basis_Proxy_Sandbox", "economics_market_raw"):
        assert rows[sheet_name]["classification"] == "optional_sector_pack_sheet"
        assert rows[sheet_name]["present_in_standard_shell"] is False


def test_visible_qa_surface_lifecycle_uses_canonical_issue_ledger_views() -> None:
    payload = json.loads(LIFECYCLE_JSON.read_text(encoding="utf-8"))
    surfaces = {row["sheet_name"]: row for row in payload["visible_qa_surfaces"]}

    assert set(surfaces) == {"QA_Log", "Needs_Review", "QA_Checks"}
    assert surfaces["QA_Log"]["source_of_data"] == "canonical issue-ledger summaries"
    assert "visibility_disposition=needs_review" in surfaces["Needs_Review"]["source_of_data"]
    assert surfaces["QA_Checks"]["source_of_data"] == "canonical issue-ledger rule aggregates"
    assert all("explicit overflow only" in row["validation_rules"] for row in surfaces.values())
