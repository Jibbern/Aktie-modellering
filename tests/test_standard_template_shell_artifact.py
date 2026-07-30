from __future__ import annotations

from datetime import date, datetime
import importlib.util
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter, range_boundaries

from scripts.build_standard_template_shell_neutrality_audit import (
    APPROVED_GENERIC_PRODUCT_LABELS,
)


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "standard_stock_model_template.xlsx"
ANF_LAB = ROOT / "templates" / "lab" / "ANF_template_lab.xlsx"
VISUAL_GAP_AUDIT_JSON = ROOT / "docs" / "standard_template_shell_visual_gap_audit.json"
VISUAL_GAP_AUDIT_MD = ROOT / "docs" / "standard_template_shell_visual_gap_audit.md"
HIDDEN_SUPPORT_AUDIT_JSON = ROOT / "docs" / "standard_template_hidden_support_audit.json"
HIDDEN_SUPPORT_AUDIT_MD = ROOT / "docs" / "standard_template_hidden_support_audit.md"
NEUTRALITY_AUDIT_JSON = ROOT / "docs" / "standard_template_shell_neutrality_audit.json"
NEUTRALITY_AUDIT_MD = ROOT / "docs" / "standard_template_shell_neutrality_audit.md"
SHEET_INVENTORY_JSON = ROOT / "docs" / "standard_template_sheet_inventory.json"
SUPPORT_LIFECYCLE_JSON = ROOT / "docs" / "support_sheet_lifecycle_contract.json"
RICH_VISIBLE_SHEETS = [
    "SUMMARY",
    "Valuation",
    "BS_Segments",
    "Operating_Drivers",
    "{ticker}_Investment_Case",
    "Quarter_Notes_UI",
    "Promise_Progress_UI",
]
COMPANY_SPECIFIC_TERMS = (
    "ANF",
    "Abercrombie",
    "Hollister",
    "Pitney Bowes",
    "Presort",
    "SendTech",
    "Green Plains",
    "45Z",
    "RIN",
    "crush margin",
)
FIXED_DIMENSION_MEMBERS = ("Americas", "EMEA", "APAC")
FIXED_SECTOR_LABEL_REGEXES = (
    r"\bclosures?\b",
    r"\bopenings?\b",
    r"\bremodels?\b",
    r"\bstores?\s*/\s*buybacks?\b",
    r"\bstores?\s*/\s*real estate\b",
    r"\bfranchise stores?\b",
    r"\bowned stores?\b",
    r"\btariffs?\b",
    r"\bERP\b",
    r"\bfreight tailwind\b",
    r"\bmarketing headwind\b",
    r"\bnet sales growth\b",
    r"\badjusted EPS\b",
    r"\bshare repurchases?\b",
    r"\breal estate activity\b",
)
SIGNAL_FILL_COLORS = {
    "002F80ED",
    "006FA8DC",
    "009BD3F5",
    "00A63A00",
    "00D55E00",
    "00E69F00",
    "00009E73",
    "0066C2A5",
    "0056B4E9",
    "00CC79A7",
}
MODULE_MANIFEST_PAYLOAD = json.loads((ROOT / "docs" / "workbook_module_manifest.json").read_text(encoding="utf-8"))
ALLOWED_HIDDEN_SHELL_SHEETS = {
    sheet["sheet"]
    for module in MODULE_MANIFEST_PAYLOAD["modules"]
    for sheet in module["sheets"]
    if sheet["role"] != "visible_product"
}
REQUIRED_SUPPORT_SHELL_SHEETS = set(ALLOWED_HIDDEN_SHELL_SHEETS)
FORMULA_OUTPUT_SUPPORT_SHEETS = {
    sheet["sheet"]
    for module in MODULE_MANIFEST_PAYLOAD["modules"]
    for sheet in module["sheets"]
    if sheet.get("data_surface") == "formula_output" or sheet.get("formula_owned_ranges")
}
HEADERED_DEBT_PRODUCT_SHELLS = {
    "Debt_Profile",
    "Debt_Credit_Notes",
    "Debt_Maturity_Ladder",
}
REQUIRED_PROMISE_ANNUAL_HEADERS = [
    "Metric",
    "Initial guide",
    "Q1 update",
    "Q2 update",
    "Q3 update",
    "Q4 update",
    "Actual",
    "Status",
    "Notes/source",
]
REQUIRED_PROMISE_REVISION_HEADERS = [
    "Metric",
    "Previous guide",
    "New/current guide",
    "Change type",
    "Actual",
    "Progress / run-rate",
    "Status",
    "Horizon",
    "Stated in",
    "Source date",
    "Source / note",
]
PROMISE_ANNUAL_HEADER_ROWS = (12, 23, 29, 34)
PROMISE_REVISION_HEADER_ROWS = (60, 70, 77, 85, 91, 98)
NEUTRAL_BLANK_FILLS = {"", "00000000", "00FFFFFF", "FFFFFFFF"}
GRAY_BLANK_FILLS = {"00DDDDDD", "00D9D9D9", "FFD9D9D9", "FFDDDDDD"}
VALUATION_GUIDANCE_SIDECAR_HEADERS = {
    "O7": "Current guidance",
    "O8": "Metric",
    "Q8": "Stated in",
    "R8": "Applies to",
    "S8": "Guidance",
    "X8": "Unit",
    "Y8": "Published",
    "Z8": "Evidence",
    "AA8": "Role / source status",
    "O27": "Historical guidance",
    "O28": "Metric",
    "Q28": "Stated in",
    "R28": "Applies to",
    "S28": "Guidance",
    "X28": "Unit",
    "Y28": "Published",
    "Z28": "Evidence",
    "AA28": "Role / source status",
    "O48": "Thesis / debate evidence",
    "O49": "Typed evidence only; unresolved synthesis remains explicit.",
    "O50": "Item",
    "Q50": "Evidence",
    "X50": "Review state",
    "Z50": "Source key",
    "O63": "Output",
    "U63": "Value",
    "X63": "Interpretation",
}
VALUATION_STRUCTURAL_HEADERS = {
    "B138": "Summary",
    "F138": "Score",
    "G138": "State",
    "H138": "As of period",
}
VALUATION_BLUE_SECTION_HEADERS = {
    "O7",
    "O27",
    "O48",
    "A122",
    "A137",
    "N137",
    "B192",
}
VALUATION_BLUE_SECTION_HEADER_RANGES = (
    "A122:N122",
    "A152:M152",
    "B192:S192",
)
STATUS_OUTPUT_FILL_COLORS = {
    "00D9EAF7",
    "00F2F2F2",
    "00F4CCCC",
    "00FFF2CC",
}
SECTION_BLUE = "006FA8DC"
HEADER_BLUE = "00EAF3FB"
OPERATING_DRIVER_SHEET_HEADERS = {
    "A12": "Topic",
    "B12": "Current read",
    "H12": "Source / use",
    "A19": "Horizon",
    "B19": "Stated in",
    "C19": "Commentary",
}
def _load_validator():
    path = ROOT / "scripts" / "validate_standard_template_shell.py"
    spec = importlib.util.spec_from_file_location("validate_standard_template_shell", path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _binding_payload() -> dict:
    import json

    return json.loads((ROOT / "docs" / "workbook_binding_map.json").read_text(encoding="utf-8"))


def _manifest_payload() -> dict:
    import json

    return json.loads((ROOT / "docs" / "standard_template_shell_manifest.json").read_text(encoding="utf-8"))


def _sheet_name(template: str, ticker: str = "ANF") -> str:
    return template.replace("{ticker}", ticker)


def _nonempty_count(ws, target: str) -> int:
    min_col, min_row, max_col, max_row = range_boundaries(target)
    count = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value not in (None, ""):
                count += 1
    return count


def _coord_in_ranges(coord: str, ranges: tuple[str, ...]) -> bool:
    row_idx, col_idx = coordinate_to_tuple(coord)
    for range_ref in ranges:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        if min_row <= row_idx <= max_row and min_col <= col_idx <= max_col:
            return True
    return False


def _sheet_counts(wb, sheet: str) -> tuple[int, int, int]:
    ws = wb[sheet]
    nonempty = 0
    formulas = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                nonempty += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
    return len(ws.merged_cells.ranges), nonempty, formulas


def test_standard_template_shell_artifact_exists() -> None:
    assert TEMPLATE.exists()
    assert TEMPLATE.suffix == ".xlsx"


def test_standard_template_uses_one_shot_full_calculation_metadata() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=True)
    try:
        assert wb.calculation.calcMode == "auto"
        assert wb.calculation.fullCalcOnLoad is True
        assert wb.calculation.forceFullCalc is False
    finally:
        wb.close()


def test_standard_template_shell_validation_passes() -> None:
    validator = _load_validator()
    report = validator.validate_shell(template_path=TEMPLATE)

    assert report["status"] == "PASS", report
    assert report["issue_count"] == 0
    json.dumps(report)


def test_filled_ticker_sheet_resolution_covers_visible_and_data_support_sheets() -> None:
    validator = _load_validator()
    workbook = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        workbook["{ticker}_Investment_Case"].title = "TEST_Investment_Case"
        workbook["{ticker}_Investment_Case_Data"].title = "TEST_Investment_Case_Data"

        assert validator._filled_ticker_sheet_names(workbook) == {
            "{ticker}_Investment_Case": "TEST_Investment_Case",
            "{ticker}_Investment_Case_Data": "TEST_Investment_Case_Data",
        }
        assert validator._workbook_sheet_name(
            workbook,
            "{ticker}_Investment_Case_Data",
            allow_filled_values=True,
        ) == "TEST_Investment_Case_Data"
    finally:
        workbook.close()


def test_standard_template_shell_is_rich_visual_shell_not_wireframe() -> None:
    template_wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    lab_wb = load_workbook(ANF_LAB, data_only=False, read_only=False)
    try:
        too_sparse: list[str] = []
        for sheet in RICH_VISIBLE_SHEETS:
            template_sheet = sheet
            lab_sheet = _sheet_name(sheet)
            template_merges, template_nonempty, template_formulas = _sheet_counts(template_wb, template_sheet)
            lab_merges, lab_nonempty, lab_formulas = _sheet_counts(lab_wb, lab_sheet)
            sheet_contract = next(row for row in _manifest_payload()["sheets"] if row["sheet"] == sheet)
            merge_floor_ratio = float(sheet_contract.get("rich_shell_lab_merge_floor_ratio", 0.55))
            if template_merges < max(1, int(lab_merges * merge_floor_ratio)):
                too_sparse.append(f"{sheet} merges {template_merges} < {lab_merges}")
            if template_nonempty < max(5, int(lab_nonempty * 0.03)):
                too_sparse.append(f"{sheet} nonempty {template_nonempty} < {lab_nonempty}")
            if lab_formulas and template_formulas < max(1, int(lab_formulas * 0.04)):
                too_sparse.append(f"{sheet} formulas {template_formulas} < {lab_formulas}")

        assert too_sparse == []
    finally:
        template_wb.close()
        lab_wb.close()


def test_standard_template_shell_binding_targets_are_blank() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        offenders: list[str] = []
        for binding in _binding_payload()["bindings"]:
            if not binding["writable"]:
                continue
            ws = wb[binding["sheet"]]
            count = _nonempty_count(ws, binding["target"])
            if count:
                offenders.append(f"{binding['binding_id']} {binding['sheet']}!{binding['target']} nonempty={count}")

        assert offenders == []
    finally:
        wb.close()


def test_standard_template_shell_has_no_company_specific_text_in_writable_zones() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        offenders: list[str] = []
        for sheet in _manifest_payload()["sheets"]:
            ws = wb[sheet["sheet"]]
            for zone in sheet["writable_zones"]:
                min_col, min_row, max_col, max_row = range_boundaries(zone["target"])
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        value = str(cell.value or "")
                        for term in COMPANY_SPECIFIC_TERMS:
                            if term.lower() in value.lower():
                                offenders.append(f"{sheet['sheet']}!{cell.coordinate}: {value}")

        assert offenders == []
    finally:
        wb.close()


def test_standard_template_shell_has_no_stale_qa_excel_tables() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        offenders = {
            sheet_name: list(wb[sheet_name].tables.keys())
            for sheet_name in ("QA_Log", "Needs_Review", "QA_Checks")
            if wb[sheet_name].tables
        }

        assert offenders == {}
    finally:
        wb.close()


def test_standard_template_shell_clears_representative_source_value_leaks() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        representative_source_cells = {
            "SUMMARY": ["A3", "A5", "A7", "B28"],
            "Valuation": ["B6", "B9", "M18", "S10", "AA14"],
            "BS_Segments": ["B7", "B47", "I49", "A62"],
            "Operating_Drivers": ["B6", "H6"],
            "{ticker}_Investment_Case": ["A226", "B230", "F231", "H235", "I240"],
            "Quarter_Notes_UI": ["B3", "B4", "B5", "B6"],
            "Promise_Progress_UI": ["B5", "C5", "G5", "C7"],
            "QA_Log": ["A2", "D1001", "J1001"],
            "Needs_Review": ["A2", "G2"],
            "QA_Checks": ["A2", "E2"],
        }
        offenders: list[str] = []
        for sheet_name, cells in representative_source_cells.items():
            ws = wb[sheet_name]
            for coord in cells:
                value = ws[coord].value
                if value not in (None, ""):
                    offenders.append(f"{sheet_name}!{coord}={value!r}")

        assert offenders == []
        assert wb["Valuation"]["O7"].value == "Current guidance"
    finally:
        wb.close()


def test_standard_template_shell_has_no_valuation_numeric_or_date_constants() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Valuation"]
        offenders = [
            f"Valuation!{cell.coordinate}={cell.value!r}"
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, (int, float, date, datetime))
        ]

        assert offenders == []
    finally:
        wb.close()


def test_standard_template_shell_clears_history_evidence_and_fixed_period_labels() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        quarter_notes = wb["Quarter_Notes_UI"]
        allowed_history_text = {
            "quarter read",
            "model read",
            "what changed",
            "watch next",
            "key caveat",
            "key developments",
            "theme",
            "what happened",
            "why it matters",
            "model / valuation implication",
            "source / confidence",
            "guidance / promise interpretation",
            "promise / guidance item",
            "read",
            "actual / progress interpretation",
            "status / caveat",
            "source",
            "model mapping / double-count guardrails",
            "driver",
            "model treatment",
            "double-count guardrail",
            "linked sheet / metric",
        }
        offenders: list[str] = []
        for row in quarter_notes.iter_rows(min_row=16):
            for cell in row:
                value = cell.value
                if value in (None, "") or (isinstance(value, str) and value.startswith("=")):
                    continue
                text = str(value).strip()
                if text.casefold() in allowed_history_text:
                    continue
                if cell.column == 1 and re.fullmatch(r"Historical quarter notes \d+", text, re.I):
                    continue
                offenders.append(f"Quarter_Notes_UI!{cell.coordinate}: {text}")

        assert offenders == []
        assert quarter_notes["C42"].value is None
        assert wb["Operating_Drivers"]["A14"].value == "Current guidance period"
        assert wb["Promise_Progress_UI"]["A11"].value == "Guidance progression - period block 1"
        assert wb["{ticker}_Investment_Case"]["A167"].value == "Calculation Details"
        assert wb["{ticker}_Investment_Case"]["A170"].value == "Calculation"
        assert wb["{ticker}_Investment_Case"]["A212"].value is None
        assert all(
            wb["{ticker}_Investment_Case"].row_dimensions[row].hidden
            for row in range(226, 241)
        )
    finally:
        wb.close()


def test_visible_shell_contains_no_internal_slot_placeholders() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        offenders: list[str] = []
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    text = cell.value.strip()
                    if re.search(r"\[[^\]]*(?:slot|dimension member|quality of earnings item)[^\]]*\]", text, re.I):
                        offenders.append(f"{ws.title}!{cell.coordinate}: {text}")

        assert offenders == []
    finally:
        wb.close()


def test_valuation_lower_runtime_value_status_and_date_cells_are_blank() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Valuation"]
        runtime_constant_ranges = (
            "D194:D216",
            "E236:E240",
            "D247:D250",
            "E253:E256",
            "L248:S250",
        )
        offenders: list[str] = []
        for range_ref in runtime_constant_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(range_ref)
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    value = cell.value
                    if value in (None, ""):
                        continue
                    if isinstance(value, str) and value.startswith("="):
                        continue
                    offenders.append(f"Valuation!{cell.coordinate}={value!r}")

        assert ws["D195"].value in (None, "")
        assert offenders == []
    finally:
        wb.close()


def test_standard_template_shell_has_no_visible_signal_leakage_in_valuation_outputs() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Valuation"]
        offenders: list[str] = []
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
            for cell in row:
                if cell.coordinate in VALUATION_BLUE_SECTION_HEADERS or _coord_in_ranges(cell.coordinate, VALUATION_BLUE_SECTION_HEADER_RANGES):
                    continue
                fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == "rgb" else ""
                if fill in SIGNAL_FILL_COLORS:
                    offenders.append(f"Valuation!{cell.coordinate} fill={fill} value={cell.value!r}")

        assert offenders == []
    finally:
        wb.close()


def test_standard_template_shell_has_no_red_green_status_outputs() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Valuation"]
        status_terms = {"PASS", "WARN", "FAIL", "N/A"}
        offenders: list[str] = []
        for row in ws.iter_rows(min_row=170, max_row=188, min_col=2, max_col=13):
            for cell in row:
                value = str(cell.value or "").strip()
                if value in status_terms or any(term in value for term in ("CFO/NI", "FCF TTM", "Net debt YoY", "Shares YoY")):
                    offenders.append(f"Valuation!{cell.coordinate}={value!r}")

        assert offenders == []
    finally:
        wb.close()


def test_valuation_preserves_standard_structural_headers_and_section_fills() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Valuation"]
        header_offenders = {
            coord: ws[coord].value
            for coord, expected in {**VALUATION_GUIDANCE_SIDECAR_HEADERS, **VALUATION_STRUCTURAL_HEADERS}.items()
            if ws[coord].value != expected
        }
        fill_offenders = {
            coord: ws[coord].fill.fgColor.rgb if ws[coord].fill and ws[coord].fill.fgColor.type == "rgb" else ""
            for coord in VALUATION_BLUE_SECTION_HEADERS
            if (ws[coord].fill.fgColor.rgb if ws[coord].fill and ws[coord].fill.fgColor.type == "rgb" else "") != "006FA8DC"
        }

        assert header_offenders == {}
        assert fill_offenders == {}
    finally:
        wb.close()


def test_valuation_lower_blocks_preserve_template_bands_merges_and_font_sizes() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Valuation"]
        merged_ranges = {str(range_ref) for range_ref in ws.merged_cells.ranges}

        assert "A122:M122" in merged_ranges
        assert ws["A122"].fill.fgColor.rgb == SECTION_BLUE
        assert ws["A122"].font.sz == 12

        assert "F123:M123" in merged_ranges
        assert "H138:I138" in merged_ranges
        assert [ws.cell(123, column).value for column in range(1, 7)] == [
            "Metric",
            "Value",
            "Unit",
            "As of",
            "Status",
            "Evidence / lineage",
        ]
        for coord in ("A123", "B123", "C123", "D123", "E123", "F123"):
            assert ws[coord].fill.fgColor.rgb == HEADER_BLUE
            assert ws[coord].font.bold is True

        assert [ws.cell(row, 1).value for row in range(124, 132)] == [
            "Cash",
            "Revolver availability",
            "Total liquidity",
            "Operating lease liabilities",
            "Core debt",
            "Net debt",
            "Net leverage",
            "Maturity detail",
        ]
        assert ws["E131"].value == "unavailable"
        assert ws["F131"].value == "No maturity schedule available."
        for row in range(124, 131):
            assert ws[f"B{row}"].value is None
            assert ws[f"D{row}"].value is None
            assert ws[f"E{row}"].value is None
            assert ws[f"F{row}"].value is None
        for row in range(132, 137):
            for column in range(1, 14):
                assert ws.cell(row, column).value is None
                assert ws.cell(row, column).protection.locked is True

        for coord in ("B138", "F138", "G138", "H138"):
            assert ws[coord].font.sz == 12
            assert ws[coord].fill.fgColor.rgb == HEADER_BLUE

        assert ws["B192"].value == "Valuation"
        assert ws["B192"].fill.fgColor.rgb == SECTION_BLUE
        assert ws["B192"].font.sz == 18
    finally:
        wb.close()


def test_valuation_blank_status_and_value_cells_are_visually_neutral() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Valuation"]
        offenders: list[str] = []
        for range_ref in ("B170:I188", "U51:U62"):
            min_col, min_row, max_col, max_row = range_boundaries(range_ref)
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    if cell.value not in (None, ""):
                        continue
                    fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == "rgb" else ""
                    if fill not in NEUTRAL_BLANK_FILLS:
                        offenders.append(f"Valuation!{cell.coordinate} fill={fill}")

        assert offenders == []
    finally:
        wb.close()


def test_promise_progress_ui_preserves_standard_guidance_revision_columns() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Promise_Progress_UI"]
        annual_headers = {
            row_idx: [ws.cell(row_idx, col).value for col in range(1, 10)]
            for row_idx in PROMISE_ANNUAL_HEADER_ROWS
        }
        revision_headers = {
            row_idx: [ws.cell(row_idx, col).value for col in range(1, 12)]
            for row_idx in PROMISE_REVISION_HEADER_ROWS
        }

        assert annual_headers == {row_idx: REQUIRED_PROMISE_ANNUAL_HEADERS for row_idx in PROMISE_ANNUAL_HEADER_ROWS}
        assert revision_headers == {row_idx: REQUIRED_PROMISE_REVISION_HEADERS for row_idx in PROMISE_REVISION_HEADER_ROWS}
    finally:
        wb.close()


def test_valuation_guidance_sidecar_preserves_repeated_standard_headers() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Valuation"]
        offenders = {
            coord: ws[coord].value
            for coord, expected in VALUATION_GUIDANCE_SIDECAR_HEADERS.items()
            if ws[coord].value != expected
        }

        assert offenders == {}
    finally:
        wb.close()


def test_operating_drivers_sheet_preserves_standard_subheaders() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Operating_Drivers"]
        offenders = {
            coord: ws[coord].value
            for coord, expected in OPERATING_DRIVER_SHEET_HEADERS.items()
            if ws[coord].value != expected
        }

        assert offenders == {}
    finally:
        wb.close()


def test_long_narrative_zones_are_wrapped_and_resized() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        summary = wb["SUMMARY"]
        assert all(summary.row_dimensions[row].height == 36.0 for row in range(17, 22))
        assert all(summary.row_dimensions[row].height == 42.0 for row in (23, 24))

        drivers = wb["Operating_Drivers"]
        for row_idx in (6, 7, 8, 9, 14, 15):
            assert drivers[f"B{row_idx}"].alignment.wrap_text is True
            assert drivers.row_dimensions[row_idx].height == 42.0

        promise = wb["Promise_Progress_UI"]
        for coord in ("B19", "C19", "D19", "E19", "B67", "C67"):
            assert promise[coord].alignment.wrap_text is True
        assert promise.row_dimensions[19].height == 60.0
        assert promise.row_dimensions[67].height == 42.0
    finally:
        wb.close()


def test_focused_pass_a_bs_segments_hierarchy_and_fiscal_headers_are_exact() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["BS_Segments"]
        assert ws["A1"].value == "Scale"
        assert ws["A4"].value == "Balance Sheet and Segments"
        assert ws["A50"].value == "Retail balance-sheet diagnostics"
        assert ws["A58"].value == "Quarterly segment revenue ($m)"
        assert ws["A59"].value == (
            "Geography and brand are separate analytical views; "
            "Total Company must not be added across both dimensions."
        )
        assert ws["A69"].value == "Annual segment revenue ($m)"
        assert ws["A38"].value == "Current ratio (x)"
        assert ws["A39"].value == "Quick ratio (x)"
        assert ws["A51"].value == "Inventory YoY (%)"
        assert ws["A53"].value == "Inventory growth less revenue growth (%)"

        fiscal_headers = [ws.cell(70, column) for column in range(2, 10)]
        assert all(cell.font.sz == 13 for cell in fiscal_headers)
        assert all(cell.alignment.horizontal == "center" for cell in fiscal_headers)
        assert len({cell.fill.fgColor.rgb for cell in fiscal_headers}) == 1
        assert len({cell.borderId if hasattr(cell, "borderId") else cell._style.borderId for cell in fiscal_headers}) == 1
        assert all(ws.row_dimensions[row].height == 18.0 for row in range(76, 79))
    finally:
        wb.close()


def test_standard_template_contains_no_red_number_format_or_font_rule() -> None:
    import zipfile

    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        offenders = [
            f"{ws.title}!{cell.coordinate}"
            for ws in wb.worksheets
            for cell in ws._cells.values()
            if "[red]" in str(cell.number_format).casefold()
        ]
        red_fonts = [
            f"{ws.title}!{cell.coordinate}"
            for ws in wb.worksheets
            for cell in ws._cells.values()
            if cell.font.color is not None
            and cell.font.color.type == "rgb"
            and str(cell.font.color.rgb).upper() in {"FFFF0000", "00FF0000"}
        ]
        assert offenders == []
        assert red_fonts == []
    finally:
        wb.close()

    with zipfile.ZipFile(TEMPLATE) as archive:
        styles_xml = archive.read("xl/styles.xml").decode("utf-8")
    assert "[red]" not in styles_xml.casefold()


def test_debt_product_shell_layout_matches_bounded_manifest_metadata() -> None:
    expected = {
        "Debt_Profile": {
            "widths": [16, 34, 24, 12, 9, 13, 16, 22, 30, 44],
            "header": 32,
            "body": 32,
            "wrap": {"B", "C", "H", "I", "J"},
                "zoom": 110,
            "last_row": 16,
        },
        "Revolver_History": {
            "widths": [12, 13, 18, 12, 12, 10, 13, 13, 13, 13, 11, 16, 12, 20, 24, 44],
            "header": 36,
            "body": 32,
            "wrap": {"N", "O", "P"},
                "zoom": 110,
            "last_row": 15,
        },
        "Leverage_Liquidity": {
            "widths": [11, 11, 13, 11, 14, 11, 14, 14, 12, 12, 13, 22, 38, 52],
            "header": 36,
            "body": 48,
            "wrap": {"L", "M", "N"},
                "zoom": 110,
            "last_row": 15,
        },
        "Debt_Credit_Notes": {
            "widths": [34, 22, 12, 13, 64, 14, 38, 20],
            "header": 34,
            "body": 48,
            "wrap": {"A", "E", "G", "H"},
                "zoom": 110,
            "last_row": 9,
        },
    }
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        for sheet_name, contract in expected.items():
            ws = wb[sheet_name]
            widths = [ws.column_dimensions[get_column_letter(index)].width for index in range(1, len(contract["widths"]) + 1)]
            assert widths == contract["widths"]
            assert ws.row_dimensions[3].height == contract["header"]
            assert all(ws.row_dimensions[row].height == contract["body"] for row in range(4, contract["last_row"] + 1))
            assert all(ws[f"{column}4"].alignment.wrap_text is True for column in contract["wrap"])
            assert ws.freeze_panes == "A4"
            assert ws.sheet_view.zoomScale == contract["zoom"]
            assert not ws.merged_cells.ranges
    finally:
        wb.close()


def test_promise_progress_status_column_preserves_neutral_zebra_fill_continuity() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Promise_Progress_UI"]
        for min_row, max_row in ((13, 20), (24, 27), (30, 32), (35, 36)):
            for row in range(min_row, max_row + 1):
                assert ws[f"H{row}"]._style.fillId == ws[f"G{row}"]._style.fillId
    finally:
        wb.close()


def test_operating_drivers_title_is_on_row_1_and_only_top_row_is_frozen() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        ws = wb["Operating_Drivers"]
        merged_ranges = {str(range_ref) for range_ref in ws.merged_cells.ranges}

        assert ws.freeze_panes == "A2"
        assert "A1:N1" in merged_ranges
        assert "A2:N2" not in merged_ranges
        assert ws["A1"].value == "Operating Drivers"
        assert ws["A2"].value in (None, "")
        assert ws["A1"].fill.fgColor.rgb == SECTION_BLUE
        assert ws["A1"].font.sz == 15
        assert ws["A1"].font.bold is True
        assert ws["A1"].alignment.horizontal == "center"
        assert ws["A1"].alignment.vertical == "center"
        assert ws.row_dimensions[1].height == 24
    finally:
        wb.close()


def test_standard_template_shell_required_support_sheets_are_neutral_shells() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        hidden_sheets = {ws.title for ws in wb.worksheets if ws.sheet_state != "visible"}

        assert REQUIRED_SUPPORT_SHELL_SHEETS <= hidden_sheets
        assert hidden_sheets <= ALLOWED_HIDDEN_SHELL_SHEETS

        for sheet_name in REQUIRED_SUPPORT_SHELL_SHEETS:
            ws = wb[sheet_name]
            assert ws.sheet_state != "visible"
            assert any(ws.cell(1, col).value not in (None, "") for col in range(1, min(ws.max_column, 12) + 1))
            if sheet_name in FORMULA_OUTPUT_SUPPORT_SHEETS:
                formulas = [
                    cell
                    for row in ws.iter_rows(min_row=2)
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                ]
                assert formulas
                assert all(cell.protection.locked for cell in formulas)
            elif sheet_name in HEADERED_DEBT_PRODUCT_SHELLS:
                assert _nonempty_count(ws, f"A2:{ws.cell(2, max(ws.max_column, 1)).coordinate}") == 0
                assert any(ws.cell(3, column).value not in (None, "") for column in range(1, ws.max_column + 1))
                if ws.max_row >= 4:
                    assert _nonempty_count(
                        ws,
                        f"A4:{ws.cell(ws.max_row, max(ws.max_column, 1)).coordinate}",
                    ) == 0
            else:
                assert _nonempty_count(ws, f"A2:{ws.cell(max(ws.max_row, 2), max(ws.max_column, 1)).coordinate}") == 0
    finally:
        wb.close()


def test_standard_template_shell_has_no_fixed_sector_or_dimension_rows() -> None:
    import re

    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        offenders: list[str] = []
        patterns = [
            *(re.compile(r"\b" + re.escape(term) + r"\b", re.I) for term in FIXED_DIMENSION_MEMBERS),
            *(re.compile(pattern, re.I) for pattern in FIXED_SECTOR_LABEL_REGEXES),
        ]
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str) or cell.value.startswith("="):
                        continue
                    value = cell.value
                    if value.strip().lower() in APPROVED_GENERIC_PRODUCT_LABELS:
                        continue
                    for pattern in patterns:
                        if pattern.search(value):
                            offenders.append(f"{ws.title}!{cell.coordinate}={value!r}")

        assert offenders == []
    finally:
        wb.close()


def test_standard_template_shell_blank_writable_zones_have_no_data_like_fills() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        offenders: list[str] = []
        for sheet in _manifest_payload()["sheets"]:
            ws = wb[sheet["sheet"]]
            for zone in sheet["writable_zones"]:
                min_col, min_row, max_col, max_row = range_boundaries(zone["target"])
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        if cell.value not in (None, ""):
                            continue
                        fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == "rgb" else ""
                        if fill in SIGNAL_FILL_COLORS or fill in GRAY_BLANK_FILLS:
                            offenders.append(f"{sheet['sheet']}!{cell.coordinate} fill={fill}")

        assert offenders == []
    finally:
        wb.close()


def test_standard_template_shell_visible_blank_cells_have_no_gray_data_fills() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        offenders: list[str] = []
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value not in (None, ""):
                        continue
                    fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == "rgb" else ""
                    if fill in GRAY_BLANK_FILLS:
                        offenders.append(f"{ws.title}!{cell.coordinate} fill={fill}")

        assert offenders == []
    finally:
        wb.close()


def test_standard_template_shell_pass_does_not_create_gtx_workbook() -> None:
    data_root = ROOT.parent / "StockModelData"
    forbidden = [
        data_root / "outputs" / "stress_tests" / "GTX_new_ticker_engine" / "GTX_model.xlsx",
        data_root / "outputs" / "Excel stock models" / "GTX_model.xlsx",
        data_root / "outputs" / "Excel stock models" / "GTX_model.xlsm",
    ]

    assert [str(path) for path in forbidden if path.exists()] == []


def test_standard_template_visual_gap_audit_covers_all_visible_sheets() -> None:
    import json

    assert VISUAL_GAP_AUDIT_JSON.exists()
    assert VISUAL_GAP_AUDIT_MD.exists()

    payload = json.loads(VISUAL_GAP_AUDIT_JSON.read_text(encoding="utf-8"))
    sheet_reports = payload["sheet_reports"]
    assert {report["sheet"] for report in sheet_reports} >= set(_manifest_payload()["visible_sheet_order"])

    for report in sheet_reports:
        assert report["preview_mode"] == "openpyxl_static_not_excel_com"
        assert "used_range" in report
        assert "non_empty_cells" in report["source_lab"]
        assert "non_empty_cells" in report["standard_shell"]
        assert "static_template_label_count" in report["standard_shell"]
        assert "row_label_count" in report["standard_shell"]
        assert "formula_count" in report["standard_shell"]
        assert "merge_count" in report["standard_shell"]
        assert "blank_writable_cells" in report
        assert isinstance(report["visually_complete"], bool)


def test_standard_template_shell_hidden_package_is_neutral() -> None:
    import re

    term_re = re.compile("|".join(r"\b" + re.escape(term) + r"\b" for term in COMPANY_SPECIFIC_TERMS), re.I)
    filename_re = re.compile(r"\b(anf|pbi|gpre|gtx)[-_][^\s]*\.(htm|html|pdf|xlsx|xls)\b", re.I)
    wb = load_workbook(TEMPLATE, data_only=False, read_only=False)
    try:
        hidden_sheets = {ws.title for ws in wb.worksheets if ws.sheet_state != "visible"}
        assert hidden_sheets <= ALLOWED_HIDDEN_SHELL_SHEETS

        offenders: list[str] = []
        for ws in wb.worksheets:
            if ws.sheet_state == "visible":
                continue
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if value in (None, ""):
                        continue
                    text = str(value)
                    if term_re.search(text) or filename_re.search(text):
                        offenders.append(f"{ws.title}!{cell.coordinate}={text[:120]}")

        assert offenders == []
    finally:
        wb.close()


def test_standard_template_shell_package_has_no_source_specific_xml_leakage() -> None:
    validator = _load_validator()

    assert validator._package_source_leakage_parts(TEMPLATE) == []


def test_standard_template_hidden_support_audit_covers_deleted_and_retained_sheets() -> None:
    import json

    assert HIDDEN_SUPPORT_AUDIT_JSON.exists()
    assert HIDDEN_SUPPORT_AUDIT_MD.exists()

    payload = json.loads(HIDDEN_SUPPORT_AUDIT_JSON.read_text(encoding="utf-8"))
    assert payload["package_dependency_check"]["missing_defined_name_sheets"] == []
    assert payload["package_dependency_check"]["missing_visible_formula_sheets"] == []
    assert payload["post_neutralization_summary"]["company_source_leakage_cells"] == 0

    rows_by_sheet = {row["sheet_name"]: row for row in payload["hidden_support_sheets"]}
    assert "Hidden_Value_Flags" in rows_by_sheet
    assert rows_by_sheet["Hidden_Value_Flags"]["present_in_shell"] is True
    assert rows_by_sheet["Hidden_Value_Flags"]["classification"] == "keep_formula_dependency"

    deleted = [row for row in payload["hidden_support_sheets"] if row["classification"] == "delete_from_shell"]
    assert deleted
    assert all(row["present_in_shell"] is False for row in deleted)


def test_standard_template_neutrality_audit_has_no_remaining_non_neutral_items() -> None:
    import json

    assert NEUTRALITY_AUDIT_JSON.exists()
    assert NEUTRALITY_AUDIT_MD.exists()

    payload = json.loads(NEUTRALITY_AUDIT_JSON.read_text(encoding="utf-8"))
    summary = payload["post_neutrality_summary"]
    assert summary["company_specific_value_count"] == 0
    assert summary["company_specific_text_count"] == 0
    assert summary["sector_specific_label_count"] == 0
    assert summary["fixed_dimension_member_count"] == 0
    assert summary["source_specific_text_count"] == 0
    assert summary["valuation_numeric_constant_count"] == 0
    assert summary["signal_fill_without_value_count"] == 0
    assert summary["valuation_signal_fill_count"] == 0
    assert summary["blank_writable_non_neutral_fill_count"] == 0
    assert summary["visible_blank_gray_fill_count"] == 0
    assert summary["red_green_status_output_count"] == 0
    assert summary["blank_status_or_value_fill_count"] == 0
    assert summary["visible_value_date_status_constant_count"] == 0
    assert summary["visible_company_source_text_count"] == 0
    assert summary["missing_required_support_shell_sheet_count"] == 0
    revised_eps_label = next(
        row
        for row in payload["cell_classifications"]
        if row["sheet"] == "{ticker}_Investment_Case" and row["cell"] == "A150"
    )
    assert revised_eps_label["value"] == "Latest-quarter adjusted EPS ($/share)"
    assert revised_eps_label["classification"] == "row_label_generic"


def test_standard_template_sheet_inventory_and_lifecycle_docs_exist() -> None:
    import json

    assert SHEET_INVENTORY_JSON.exists()
    assert SUPPORT_LIFECYCLE_JSON.exists()

    inventory = json.loads(SHEET_INVENTORY_JSON.read_text(encoding="utf-8"))
    lifecycle = json.loads(SUPPORT_LIFECYCLE_JSON.read_text(encoding="utf-8"))
    inventory_rows = {row["sheet_name"]: row for row in inventory["sheets"]}
    lifecycle_rows = {row["sheet_name"]: row for row in lifecycle["support_sheets"]}
    module_contracts = {
        sheet["sheet"]: sheet
        for module in MODULE_MANIFEST_PAYLOAD["modules"]
        for sheet in module["sheets"]
    }

    for sheet_name in REQUIRED_SUPPORT_SHELL_SHEETS:
        expected_classification = {
            "B": "required_support_shell_sheet",
            "C": "optional_module_shell_sheet",
            "E": "fixture_capacity_shell_sheet",
        }[module_contracts[sheet_name]["legacy_class"]]
        assert inventory_rows[sheet_name]["classification"] == expected_classification
        assert inventory_rows[sheet_name]["present_in_standard_shell"] is True
        assert lifecycle_rows[sheet_name]["owner"] == "frozen_shell"
        assert lifecycle_rows[sheet_name]["neutral_shell_required"] is True

    for sheet_name in ("Guidance_Raw", "Quarter_Notes_Audit", "DATA_Facts_Long"):
        assert sheet_name in lifecycle_rows
        assert lifecycle_rows[sheet_name]["owner"] == "external_normalized_json"
        assert lifecycle_rows[sheet_name]["neutral_shell_required"] is False


def test_root_default_validation_tickers_remain_standard_three() -> None:
    text = (ROOT / "pbi_xbrl" / "workbook_validation_runner.py").read_text(encoding="utf-8")

    assert 'TICKERS: Sequence[str] = ("PBI", "GPRE", "ANF")' in text
    assert '"GTX"' not in text.split("TICKERS: Sequence[str] =", 1)[1].split("\n", 1)[0]
