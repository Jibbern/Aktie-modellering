from __future__ import annotations

import json
import zipfile
from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import absolute_coordinate, quote_sheetname, range_boundaries
import pytest

from pbi_xbrl.standard_template_shell_identity import (
    SHELL_SEMANTIC_CONTRACT_VERSION,
    SHEET_VIEW_CONTRACT_IGNORED_PROPERTIES,
    SHEET_VIEW_CONTRACT_OWNED_PROPERTIES,
    compute_shell_identity,
    _canonical_cell_style,
    _canonical_formula,
    _planned_cell_values_equal,
    _quantize_dimension,
    validate_verified_shell_token,
    verify_post_fill_structural_identity,
    verify_shell_identity,
)
from pbi_xbrl.new_ticker_binding_planner import reproduce_binding_plan_snapshot
from pbi_xbrl.new_ticker_value_filler import _execute_binding_plan, _resolve_ticker_sheet


ROOT = Path(__file__).resolve().parents[1]
SHELL = ROOT / "templates" / "standard_stock_model_template.xlsx"
MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
BINDINGS = ROOT / "docs" / "workbook_binding_map.json"
PACKAGE = ROOT / "tests" / "fixtures" / "normalized_packages" / "TEST_minimal_valid.json"


def _contracts() -> tuple[dict, dict]:
    return (
        json.loads(MANIFEST.read_text(encoding="utf-8")),
        json.loads(BINDINGS.read_text(encoding="utf-8")),
    )


def test_checked_in_shell_identity_matches_manifest() -> None:
    manifest, bindings = _contracts()
    report = verify_shell_identity(SHELL, manifest=manifest, binding_payload=bindings)

    assert report["status"] == "PASS", report["issues"]


def test_shell_hash_mismatch_fails(tmp_path: Path) -> None:
    manifest, bindings = _contracts()
    copied = tmp_path / "shell.xlsx"
    copied.write_bytes(SHELL.read_bytes() + b"drift")

    report = verify_shell_identity(copied, manifest=manifest, binding_payload=bindings)

    assert report["status"] == "FAIL"
    assert any(issue["rule_id"] in {"shell_identity_unavailable", "shell_workbook_sha256_mismatch"} for issue in report["issues"])


def test_byte_identical_renamed_shell_passes_identity(tmp_path: Path) -> None:
    manifest, bindings = _contracts()
    copied = tmp_path / "renamed-but-identical.xlsx"
    copied.write_bytes(SHELL.read_bytes())

    report = verify_shell_identity(copied, manifest=manifest, binding_payload=bindings)

    assert report["status"] == "PASS", report["issues"]


def test_unsupported_semantic_contract_version_fails() -> None:
    manifest, bindings = _contracts()
    drifted = deepcopy(manifest)
    drifted["semantic_contract_version"] = "9.9.9"
    drifted["shell_identity"] = compute_shell_identity(
        SHELL,
        manifest=drifted,
        binding_payload=bindings,
        semantic_contract_version="9.9.9",
    )

    report = verify_shell_identity(SHELL, manifest=drifted, binding_payload=bindings)

    assert SHELL_SEMANTIC_CONTRACT_VERSION != "9.9.9"
    assert report["status"] == "FAIL"
    assert any(issue["rule_id"] == "shell_semantic_contract_version_unsupported" for issue in report["issues"])


def test_merge_and_defined_name_drift_are_detected_in_identity_contract() -> None:
    manifest, bindings = _contracts()
    actual = compute_shell_identity(SHELL, manifest=manifest, binding_payload=bindings)
    drifted = deepcopy(manifest)
    drifted["shell_identity"] = dict(actual)
    drifted["shell_identity"]["merge_signature"] = "0" * 64
    drifted["shell_identity"]["defined_name_signature"] = "1" * 64

    report = verify_shell_identity(SHELL, manifest=drifted, binding_payload=bindings)

    rules = {issue["rule_id"] for issue in report["issues"]}
    assert "shell_merge_drift" in rules
    assert "shell_defined_name_drift" in rules


def test_sheet_view_contract_and_signature_are_manifest_owned() -> None:
    manifest, bindings = _contracts()
    contract = manifest["sheet_view_identity_contract"]
    actual = compute_shell_identity(SHELL, manifest=manifest, binding_payload=bindings)
    drifted = deepcopy(manifest)
    drifted["shell_identity"] = dict(actual)
    drifted["shell_identity"]["sheet_view_signature"] = "0" * 64

    report = verify_shell_identity(SHELL, manifest=drifted, binding_payload=bindings)

    assert tuple(contract["contract_owned"]) == SHEET_VIEW_CONTRACT_OWNED_PROPERTIES
    assert tuple(contract["intentionally_ignored"]) == SHEET_VIEW_CONTRACT_IGNORED_PROPERTIES
    assert report["status"] == "FAIL"
    assert "shell_sheet_view_drift" in {issue["rule_id"] for issue in report["issues"]}


def test_binding_target_drift_is_detected() -> None:
    manifest, bindings = _contracts()
    drifted_bindings = deepcopy(bindings)
    drifted_bindings["bindings"][0]["planner_target"] = "A999"

    report = verify_shell_identity(SHELL, manifest=manifest, binding_payload=drifted_bindings)

    assert report["status"] == "FAIL"
    assert any(issue["rule_id"] == "shell_writable_target_drift" for issue in report["issues"])


def test_binding_source_and_selector_semantic_drift_is_detected() -> None:
    manifest, bindings = _contracts()

    source_drift = deepcopy(bindings)
    latest_revenue = next(row for row in source_drift["bindings"] if row["binding_id"] == "summary_latest_revenue")
    latest_revenue["source_field"] = "net_income"
    source_report = verify_shell_identity(SHELL, manifest=manifest, binding_payload=source_drift)

    selector_drift = deepcopy(bindings)
    quarter_notes = next(row for row in selector_drift["bindings"] if row["binding_id"] == "qn_quarter_note_rows")
    quarter_notes["row_selector"]["filters"][0]["equals"] = "history"
    selector_report = verify_shell_identity(SHELL, manifest=manifest, binding_payload=selector_drift)

    version_drift = deepcopy(bindings)
    version_drift["binding_planner_contract_version"] = "9.9.9"
    version_report = verify_shell_identity(SHELL, manifest=manifest, binding_payload=version_drift)

    policy_drift = deepcopy(bindings)
    policy_drift["binding_planner_policy"]["sequential_range_dumping"] = "allowed"
    policy_report = verify_shell_identity(SHELL, manifest=manifest, binding_payload=policy_drift)

    for report in (source_report, selector_report, version_report, policy_report):
        assert report["status"] == "FAIL"
        assert any(issue["rule_id"] == "shell_binding_contract_drift" for issue in report["issues"])


def test_verified_token_is_bound_to_complete_manifest_identity() -> None:
    manifest, bindings = _contracts()
    token = verify_shell_identity(SHELL, manifest=manifest, binding_payload=bindings)

    mutations: list[tuple[dict, dict]] = []
    wrong_version = deepcopy(manifest)
    wrong_version["semantic_contract_version"] = "9.9.9"
    mutations.append((wrong_version, bindings))
    other_manifest = deepcopy(manifest)
    other_manifest["description"] = "Different executable manifest semantics."
    mutations.append((other_manifest, bindings))
    wrong_hash = deepcopy(manifest)
    wrong_hash["shell_identity"]["workbook_sha256"] = "0" * 64
    mutations.append((wrong_hash, bindings))
    wrong_merge = deepcopy(manifest)
    wrong_merge["shell_identity"]["merge_signature"] = "1" * 64
    mutations.append((wrong_merge, bindings))
    wrong_bindings = deepcopy(bindings)
    next(row for row in wrong_bindings["bindings"] if row["binding_id"] == "summary_latest_revenue")["source_field"] = "net_income"
    mutations.append((manifest, wrong_bindings))
    wrong_binding_version = deepcopy(bindings)
    wrong_binding_version["binding_planner_contract_version"] = "9.9.9"
    mutations.append((manifest, wrong_binding_version))

    for candidate_manifest, candidate_bindings in mutations:
        assert validate_verified_shell_token(
            token,
            manifest=candidate_manifest,
            binding_payload=candidate_bindings,
        )

    assert validate_verified_shell_token(
        {"status": "PASS"},  # type: ignore[arg-type]
        manifest=manifest,
        binding_payload=bindings,
    )[0]["rule_id"] == "shell_identity_not_verified"


def test_post_fill_structural_identity_allows_only_approved_value_changes_in_memory() -> None:
    manifest, bindings = _contracts()
    token = verify_shell_identity(SHELL, manifest=manifest, binding_payload=bindings)
    package = json.loads(PACKAGE.read_text(encoding="utf-8"))
    plan, _snapshot = reproduce_binding_plan_snapshot(
        package,
        manifest=manifest,
        binding_payload=bindings,
        shell_path=SHELL,
        shell_identity_report=token,
    )
    assert plan.status == "PASS"
    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        _resolve_ticker_sheet(wb, "TEST")
        _execute_binding_plan(wb, plan)
        report = verify_post_fill_structural_identity(
            wb,
            approved_shell_path=SHELL,
            manifest=manifest,
            binding_payload=bindings,
            approved_plan=plan,
            normalized_package=package,
        )
        assert report["status"] == "PASS", report["issues"]
        assert report["changed_writable_cell_count"] > 0
    finally:
        wb.close()


def test_post_fill_structural_identity_requires_plan_for_changed_values() -> None:
    manifest, bindings = _contracts()
    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        wb["SUMMARY"]["A3"] = "Unplanned value"
        report = verify_post_fill_structural_identity(
            wb,
            approved_shell_path=SHELL,
            manifest=manifest,
            binding_payload=bindings,
        )
    finally:
        wb.close()

    assert report["status"] == "FAIL"
    assert "post_fill_normalized_package_required" in {row["rule_id"] for row in report["issues"]}


def test_post_fill_rejects_fabricated_pass_plan_mapping() -> None:
    manifest, bindings = _contracts()
    package = json.loads(PACKAGE.read_text(encoding="utf-8"))
    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        wb["SUMMARY"]["A3"] = "FORGED VALUE"
        fabricated = {
            "plan_version": "1.0.0",
            "ticker": "TEST",
            "status": "PASS",
            "planned_write_count": 1,
            "planned_writes": [
                {
                    "binding_id": "summary_company_description",
                    "normalized_path": "company_profile.business_description",
                    "row_key": "scalar",
                    "target_sheet": "SUMMARY",
                    "target_cell": "A3",
                    "target_type": "text",
                    "target_role": "summary_company_description",
                    "value": "FORGED VALUE",
                    "value_type": "string",
                    "source_ref": "fabricated",
                    "capacity_used": 1,
                }
            ],
            "bindings": [],
            "issue_ledger": {},
            "shell_identity": {},
        }
        report = verify_post_fill_structural_identity(
            wb,
            approved_shell_path=SHELL,
            manifest=manifest,
            binding_payload=bindings,
            approved_plan=fabricated,
            normalized_package=package,
        )
    finally:
        wb.close()

    assert report["status"] == "FAIL"
    assert "post_fill_binding_plan_reproduction_failed" in {row["rule_id"] for row in report["issues"]}


def test_post_fill_reproduces_supplied_plan_even_when_shell_is_unchanged() -> None:
    manifest, bindings = _contracts()
    package = json.loads(PACKAGE.read_text(encoding="utf-8"))
    fabricated = {
        "plan_version": "1.0.0",
        "ticker": "TEST",
        "status": "PASS",
        "planned_write_count": 1,
        "planned_writes": [
            {
                "binding_id": "summary_company_description",
                "normalized_path": "company_profile.business_description",
                "row_key": "scalar",
                "target_sheet": "SUMMARY",
                "target_cell": "A3",
                "target_type": "text",
                "target_role": "summary_company_description",
                "value": "FORGED VALUE",
                "value_type": "string",
                "source_ref": "fabricated",
                "capacity_used": 1,
            }
        ],
        "bindings": [],
        "issue_ledger": {},
        "shell_identity": {},
    }
    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        report = verify_post_fill_structural_identity(
            wb,
            approved_shell_path=SHELL,
            manifest=manifest,
            binding_payload=bindings,
            approved_plan=fabricated,
            normalized_package=package,
        )
    finally:
        wb.close()

    assert report["changed_writable_cell_count"] == 0
    assert report["status"] == "FAIL"
    assert "post_fill_binding_plan_reproduction_failed" in {row["rule_id"] for row in report["issues"]}


def test_post_fill_rejects_unchanged_shell_when_expected_plan_values_are_missing() -> None:
    manifest, bindings = _contracts()
    package = json.loads(PACKAGE.read_text(encoding="utf-8"))
    plan, _snapshot = reproduce_binding_plan_snapshot(
        package,
        manifest=manifest,
        binding_payload=bindings,
        shell_path=SHELL,
    )
    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        report = verify_post_fill_structural_identity(
            wb,
            approved_shell_path=SHELL,
            manifest=manifest,
            binding_payload=bindings,
            approved_plan=plan,
            normalized_package=package,
        )
    finally:
        wb.close()

    assert report["changed_writable_cell_count"] == 0
    assert report["status"] == "FAIL"
    assert "post_fill_planned_value_mismatch" in {row["rule_id"] for row in report["issues"]}


def test_post_fill_structural_identity_rejects_merge_name_and_protected_cell_drift_in_memory() -> None:
    manifest, bindings = _contracts()

    merge_wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        merge_ref = str(next(iter(merge_wb["Valuation"].merged_cells.ranges)))
        merge_wb["Valuation"].unmerge_cells(merge_ref)
        merge_report = verify_post_fill_structural_identity(
            merge_wb,
            approved_shell_path=SHELL,
            manifest=manifest,
            binding_payload=bindings,
        )
    finally:
        merge_wb.close()

    name_wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        name_wb.defined_names["valuation_share_count_anchor"].attr_text = "'Valuation'!$A$103"
        name_report = verify_post_fill_structural_identity(
            name_wb,
            approved_shell_path=SHELL,
            manifest=manifest,
            binding_payload=bindings,
        )
    finally:
        name_wb.close()

    protected_wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        protected_wb["SUMMARY"]["A1"] = "Unauthorized structural change"
        protected_report = verify_post_fill_structural_identity(
            protected_wb,
            approved_shell_path=SHELL,
            manifest=manifest,
            binding_payload=bindings,
        )
    finally:
        protected_wb.close()

    assert "post_fill_merge_drift" in {row["rule_id"] for row in merge_report["issues"]}
    assert "post_fill_defined_name_drift" in {row["rule_id"] for row in name_report["issues"]}
    assert "post_fill_protected_cell_drift" in {row["rule_id"] for row in protected_report["issues"]}


@pytest.mark.parametrize(
    ("attribute", "value"),
    [
        ("zoomScale", 175),
        ("showGridLines", False),
        ("showRowColHeaders", False),
    ],
)
def test_post_fill_structural_identity_rejects_contract_owned_sheet_view_drift(
    attribute: str,
    value: object,
) -> None:
    manifest, bindings = _contracts()
    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        setattr(wb["SUMMARY"].sheet_view, attribute, value)
        report = verify_post_fill_structural_identity(
            wb,
            approved_shell_path=SHELL,
            manifest=manifest,
            binding_payload=bindings,
        )
    finally:
        wb.close()

    assert report["status"] == "FAIL"
    assert "post_fill_sheet_view_drift" in {row["rule_id"] for row in report["issues"]}


def test_post_fill_structural_identity_ignores_volatile_selection_and_scroll_state() -> None:
    manifest, bindings = _contracts()
    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        view = wb["SUMMARY"].sheet_view
        view.topLeftCell = "C9"
        selection = view.selection[0]
        selection.activeCell = "C9"
        selection.sqref = "C9"
        report = verify_post_fill_structural_identity(
            wb,
            approved_shell_path=SHELL,
            manifest=manifest,
            binding_payload=bindings,
        )
    finally:
        wb.close()

    assert report["status"] == "PASS", report["issues"]


def test_post_fill_structural_identity_rejects_unsupported_manifest_version() -> None:
    manifest, bindings = _contracts()
    drifted = deepcopy(manifest)
    drifted["semantic_contract_version"] = "9.9.9"
    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        report = verify_post_fill_structural_identity(
            wb,
            approved_shell_path=SHELL,
            manifest=drifted,
            binding_payload=bindings,
        )
    finally:
        wb.close()

    assert report["status"] == "FAIL"
    assert "shell_semantic_contract_version_unsupported" in {row["rule_id"] for row in report["issues"]}


def test_checked_in_package_uses_deterministic_zip_metadata() -> None:
    with zipfile.ZipFile(SHELL) as archive:
        assert {info.date_time for info in archive.infolist()} == {(2000, 1, 1, 0, 0, 0)}
        core = archive.read("docProps/core.xml").decode("utf-8")

    assert core.count("2000-01-01T00:00:00Z") == 2


def test_company_specific_labels_and_constant_defined_name_are_absent() -> None:
    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        assert [wb["Valuation"][f"A{row}"].value for row in range(36, 41)] == [
            "Net income",
            "Net margin %",
            "Net income YoY %",
            "Net income (TTM)",
            "Net margin (TTM)",
        ]
        assert wb["SUMMARY"]["A8"].value is None
        assert wb["{ticker}_Investment_Case"]["A1"].value is None
        assert "ThesisBaseAdjEBITDA_FY" not in wb.defined_names
        assert wb["{ticker}_Investment_Case"]["A193"].value == "Business Health"
        assert wb["{ticker}_Investment_Case"]["A211"].value == "Asset Productivity / Capacity Returns"
        assert not wb["{ticker}_Investment_Case"].data_validations.dataValidation
        assert wb.defined_names["valuation_share_count_anchor"].attr_text == "'Valuation'!$A$102"
        assert wb.defined_names["investment_key_debate_anchor"].attr_text == "'{ticker}_Investment_Case'!$A$7"
    finally:
        wb.close()


def test_excel_roundtrip_canonicalization_preserves_semantic_differences() -> None:
    assert _canonical_formula("=A1+1.0") == "=A1+1"
    assert _canonical_formula("='Valuation'!A1+1.00") == "=Valuation!A1+1"
    assert _canonical_formula("=A1+2") != _canonical_formula("=A1+1")
    assert _quantize_dimension(15.001, step="0.05") == 15.0
    assert _quantize_dimension(15.04, step="0.05") == 15.05
    assert _planned_cell_values_equal(1.103709164274075, 1.1037091642740746)
    assert not _planned_cell_values_equal(1.1038, 1.1037)
    assert not _planned_cell_values_equal("1.0", 1.0)

    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        source = wb["SUMMARY"]["A1"]
        same = wb["SUMMARY"]["A2"]
        same._style = deepcopy(source._style)
        assert _canonical_cell_style(source) == _canonical_cell_style(same)
        same.fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
        assert _canonical_cell_style(source) != _canonical_cell_style(same)
    finally:
        wb.close()


def test_active_binding_defined_names_match_exact_planner_targets() -> None:
    manifest, bindings = _contracts()
    del manifest
    wb = load_workbook(SHELL, read_only=False, data_only=False)
    try:
        for binding in bindings["bindings"]:
            if not binding.get("writable") or binding.get("planning_state", "active") != "active":
                continue
            binding_id = binding["binding_id"]
            assert binding_id in wb.defined_names
            min_col, min_row, _max_col, _max_row = range_boundaries(binding.get("planner_target") or binding["target"])
            sheet_name = binding["sheet"]
            coordinate = wb[sheet_name].cell(min_row, min_col).coordinate
            assert wb.defined_names[binding_id].attr_text == f"{quote_sheetname(sheet_name)}!{absolute_coordinate(coordinate)}"
    finally:
        wb.close()
