from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any, List, Set, Tuple

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


WORKBOOK_DIR = Path(os.environ.get("STOCK_MODEL_WORKBOOK_DIR", r"C:\Users\Jibbe\Aktier\Excel stock models"))
TICKERS = ("PBI", "GPRE", "ANF")


def _load_workbook(ticker: str):
    path = WORKBOOK_DIR / f"{ticker}_model.xlsx"
    if not path.exists():
        pytest.skip(f"{path} is not available for layout polish regression tests")
    return load_workbook(path, data_only=False, read_only=False)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _row_text(ws: Worksheet, row: int, *, max_col: int = 12) -> str:
    return " ".join(_text(ws.cell(row, cc).value) for cc in range(1, max_col + 1) if _text(ws.cell(row, cc).value))


def _find_rows_containing(ws: Worksheet, needle: str) -> List[int]:
    needle_low = needle.lower()
    rows: List[int] = []
    for rr in range(1, int(ws.max_row or 0) + 1):
        if needle_low in _row_text(ws, rr, max_col=max(12, int(ws.max_column or 0))).lower():
            rows.append(rr)
    return rows


def _height(ws: Worksheet, row: int) -> float:
    return float(ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight or 15.0)


def _assert_height_close(ws: Worksheet, row: int, expected: float, *, context: str) -> None:
    actual = _height(ws, row)
    assert abs(actual - expected) <= 0.35, f"{context}: expected row height {expected}, got {actual}"


def _is_blue_section_row(ws: Worksheet, row: int) -> bool:
    fill = _text(ws.cell(row, 1).fill.fgColor.rgb).upper()
    return fill.endswith("5B9BD5") or fill.endswith("6FA8DC")


def _quarter_header_rows(ws: Worksheet) -> List[int]:
    return [
        rr
        for rr in range(1, int(ws.max_row or 0) + 1)
        if _text(ws.cell(rr, 1).value).endswith(" - Quarter Notes")
    ]


def _quarter_header_labels(ws: Worksheet) -> Set[str]:
    return {_text(ws.cell(rr, 1).value).replace(" - Quarter Notes", "") for rr in _quarter_header_rows(ws)}


def _merged_row_width(ws: Worksheet, row: int) -> int:
    for merged in ws.merged_cells.ranges:
        if merged.min_row == row and merged.max_row == row and merged.min_col == 1:
            return int(merged.max_col)
    return 1


def _history_quarter_labels(wb: Any, limit: int = 8) -> List[str]:
    ws = wb["History_Q"]
    headers = [_text(ws.cell(1, cc).value).lower() for cc in range(1, int(ws.max_column or 0) + 1)]

    label_col = None
    for name in ("fiscal_period", "fiscal label", "fiscal_label", "quarter", "period"):
        if name in headers:
            label_col = headers.index(name) + 1
            break
    fy_col = headers.index("fiscal_year") + 1 if "fiscal_year" in headers else None
    fq_col = headers.index("fiscal_quarter") + 1 if "fiscal_quarter" in headers else None

    labels: Set[str] = set()
    for rr in range(2, int(ws.max_row or 0) + 1):
        label = ""
        if label_col:
            raw = _text(ws.cell(rr, label_col).value)
            match = re.search(r"(20\d{2})-Q([1-4])", raw)
            if match:
                label = f"{match.group(1)}-Q{match.group(2)}"
        if not label and fy_col and fq_col:
            year_raw = _text(ws.cell(rr, fy_col).value)
            quarter_raw = _text(ws.cell(rr, fq_col).value)
            year_match = re.search(r"20\d{2}", year_raw)
            quarter_match = re.search(r"[1-4]", quarter_raw)
            if year_match and quarter_match:
                label = f"{year_match.group(0)}-Q{quarter_match.group(0)}"
        if label:
            labels.add(label)

    def _key(label: str) -> Tuple[int, int]:
        match = re.fullmatch(r"(20\d{2})-Q([1-4])", label)
        return (int(match.group(1)), int(match.group(2))) if match else (0, 0)

    return sorted(labels, key=_key, reverse=True)[:limit]


def test_investment_case_row_height_polish() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        ws = wb[f"{ticker}_Investment_Case"]

        snapshot_rows = _find_rows_containing(ws, "Investment Snapshot")
        assert snapshot_rows, f"{ticker}: missing Investment Snapshot section"
        start = snapshot_rows[0]
        body_rows: List[int] = []
        for rr in range(start + 1, int(ws.max_row or 0) + 1):
            first = _text(ws.cell(rr, 1).value)
            if not first:
                break
            if _is_blue_section_row(ws, rr):
                break
            body_rows.append(rr)
        assert body_rows, f"{ticker}: no Investment Snapshot body rows found"
        for rr in body_rows:
            _assert_height_close(ws, rr, 24.0, context=f"{ticker} {ws.title}!A{rr} Investment Snapshot body")

        note_rows = _find_rows_containing(ws, "Uses Investment_Case manual inputs; may differ from Valuation Thesis Bridge.")
        assert note_rows, f"{ticker}: missing Bear/Base/Bull source note row"
        for rr in note_rows:
            _assert_height_close(ws, rr, 13.5, context=f"{ticker} {ws.title}!A{rr} Bear/Base/Bull note")


def test_promise_progress_spacer_rows_are_readable() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        ws = wb["Promise_Progress_UI"]
        blank_rows = [
            rr
            for rr in range(2, int(ws.max_row or 0))
            if not _row_text(ws, rr, max_col=10)
            and _row_text(ws, rr - 1, max_col=10)
            and _row_text(ws, rr + 1, max_col=10)
        ]
        assert blank_rows, f"{ticker}: expected visible spacer rows in Promise_Progress_UI"
        for rr in blank_rows:
            _assert_height_close(ws, rr, 18.0, context=f"{ticker} Promise_Progress_UI!A{rr} spacer")


def test_quarter_notes_layout_is_wide_readable_and_covers_recent_quarters() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        ws = wb["Quarter_Notes_UI"]
        headers = _quarter_header_rows(ws)
        assert headers, f"{ticker}: Quarter_Notes_UI has no quarter headers"
        assert max(_merged_row_width(ws, rr) for rr in headers) >= 11, f"{ticker}: quarter headers should span beyond A:J"

        used_cols = {
            cc
            for row in ws.iter_rows(min_row=1, max_row=min(220, int(ws.max_row or 0)), max_col=max(12, int(ws.max_column or 0)))
            for cc, cell in enumerate(row, start=1)
            if _text(cell.value)
        }
        assert max(used_cols or {0}) >= 11, f"{ticker}: Quarter_Notes_UI should use wider narrative columns"

        body_font_sizes = [
            float(ws.cell(rr, cc).font.sz or 0)
            for rr in range(1, min(220, int(ws.max_row or 0)) + 1)
            for cc in range(1, min(12, int(ws.max_column or 0)) + 1)
            if _text(ws.cell(rr, cc).value)
            and not _text(ws.cell(rr, 1).value).endswith(" - Quarter Notes")
            and _text(ws.cell(rr, 1).value)
            not in {"Quarter read", "Key developments", "Guidance / Promise interpretation", "Model mapping / double-count guardrails"}
        ]
        assert body_font_sizes and min(body_font_sizes) >= 13.0, f"{ticker}: narrative body font should be about 14pt"

        missing_labels = set(_history_quarter_labels(wb, limit=8)) - _quarter_header_labels(ws)
        assert not missing_labels, f"{ticker}: Quarter_Notes_UI missing recent quarter blocks {sorted(missing_labels)}"

        all_text = "\n".join(_row_text(ws, rr, max_col=max(12, int(ws.max_column or 0))) for rr in range(1, int(ws.max_row or 0) + 1))
        assert "No source-backed narrative items generated for this quarter." in all_text, (
            f"{ticker}: sparse quarters should show an explicit no-information row"
        )


def test_quarter_narrative_data_is_ordered_with_audit_sheets() -> None:
    for ticker in TICKERS:
        wb = _load_workbook(ticker)
        names = list(wb.sheetnames)
        assert "Quarter_Narrative_Data" in names, f"{ticker}: missing Quarter_Narrative_Data"
        assert names.index("Quarter_Narrative_Data") > names.index("Promise_Progress_UI"), (
            f"{ticker}: Quarter_Narrative_Data should not sit among primary user-facing sheets"
        )
        if "Quarter_Notes" in names:
            assert names.index("Quarter_Narrative_Data") >= names.index("Quarter_Notes"), (
                f"{ticker}: Quarter_Narrative_Data should live near Quarter_Notes audit/data sheets"
            )
