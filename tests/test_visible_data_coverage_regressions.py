from __future__ import annotations

import datetime as dt
import os
import re
from copy import deepcopy
from dataclasses import replace
from pathlib import Path
from typing import Any

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pbi_xbrl.excel_writer_bs_segments import _should_render_carbon_equipment_liabilities
from pbi_xbrl.excel_writer_coloring import _hidden_source_comparison_metric
from pbi_xbrl.excel_writer_segment_sources import (
    _merge_quarterly_segment_packages_per_period,
    _pbi_add_corporate_reconciliation_from_release_text,
    _pbi_repair_total_reportable_segment_quarterly_totals_for_bs,
    _segment_residual_ledger_payload,
    _segment_package_ledger_merge_validation,
)
from pbi_xbrl.pipeline_orchestration import _parse_local_non_gaap_segment_rows_from_text
from pbi_xbrl.segment_normalization import (
    SEGMENT_EXACT_RESIDUAL_RULE_ID,
    SEGMENT_EXACT_ZERO_CLASSIFICATION,
    SEGMENT_RESIDUAL_LEDGER_CONTRACT_ID,
    SegmentNormalizationError,
    SegmentResidualInputFact,
    derive_exact_zero_segment_residual,
    segment_residual_input_fact_from_legacy_row,
    validate_segment_residual_ledger_payload,
)


WORKBOOK_DIR = Path(
    os.environ.get(
        "STOCK_MODEL_WORKBOOK_DIR",
        r"C:\Users\Jibbe\Aktier\StockModelData\outputs\Excel stock models",
    )
)

_QUARTER_LABEL_RE = re.compile(r"^(20\d{2})-Q([1-4])$")


def _model_path(ticker: str) -> Path:
    candidates = [WORKBOOK_DIR / f"{ticker}_model.xlsm", WORKBOOK_DIR / f"{ticker}_model.xlsx"]
    for path in candidates:
        if path.exists():
            return path
    pytest.skip(f"{ticker} model not found in {WORKBOOK_DIR}")


def _load_model(ticker: str):
    # Headless fresh-generation validation may intentionally use macro-free .xlsx
    # files generated with --skip-macro-injection. Macro-enabled .xlsm validation
    # is a separate production gate when Excel COM is available.
    return load_workbook(_model_path(ticker), data_only=True, read_only=True, keep_vba=True)


def _load_model_with_styles(ticker: str):
    return load_workbook(_model_path(ticker), data_only=True, read_only=False, keep_vba=True)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("%", ""))
    except ValueError:
        return None


def _find_row(ws: Worksheet, label: str, *, start: int = 1, end: int | None = None, col: int = 1) -> int:
    max_row = end or int(ws.max_row or 0)
    for rr in range(start, max_row + 1):
        if _text(ws.cell(rr, col).value) == label:
            return rr
    raise AssertionError(f"{ws.title}: could not find row label {label!r}")


def _find_row_contains(ws: Worksheet, label: str, *, start: int = 1, end: int | None = None) -> int:
    max_row = end or int(ws.max_row or 0)
    needle = label.lower()
    for rr in range(start, max_row + 1):
        row_text = " ".join(_text(ws.cell(rr, cc).value) for cc in range(1, int(ws.max_column or 0) + 1))
        if needle in row_text.lower():
            return rr
    raise AssertionError(f"{ws.title}: could not find row containing {label!r}")


def _quarter_col(ws: Worksheet, header_row: int, quarter: str) -> int:
    for cc in range(1, int(ws.max_column or 0) + 1):
        if _text(ws.cell(header_row, cc).value) == quarter:
            return cc
    raise AssertionError(f"{ws.title}: quarter {quarter!r} missing from row {header_row}")


def _quarter_key(label: str) -> tuple[int, int]:
    match = _QUARTER_LABEL_RE.fullmatch(_text(label))
    if not match:
        raise AssertionError(f"bad quarter label {label!r}")
    return int(match.group(1)), int(match.group(2))


def _prior_year_quarter(label: str) -> str:
    year, quarter = _quarter_key(label)
    return f"{year - 1}-Q{quarter}"


def _quarter_labels_from_header(ws: Worksheet, header_row: int) -> list[str]:
    labels: list[str] = []
    for cc in range(2, int(ws.max_column or 0) + 1):
        label = _text(ws.cell(header_row, cc).value)
        if _QUARTER_LABEL_RE.fullmatch(label):
            labels.append(label)
            continue
        if labels:
            break
    assert labels, f"{ws.title}: expected visible quarter labels on row {header_row}"
    assert len(labels) == len(set(labels)), f"{ws.title}: visible quarter labels should be unique"
    assert labels == sorted(labels, key=_quarter_key), f"{ws.title}: visible quarter labels should be ordered"
    return labels


def _history_records_by_fiscal_label(wb: Any) -> dict[str, dict[str, Any]]:
    hist = wb["History_Q"]
    headers = [_text(hist.cell(1, cc).value) for cc in range(1, int(hist.max_column or 0) + 1)]
    idx = {header: pos + 1 for pos, header in enumerate(headers) if header}
    assert {"quarter", "fiscal_label"}.issubset(idx), "History_Q should expose quarter and fiscal_label"
    records: dict[str, dict[str, Any]] = {}
    for rr in range(2, int(hist.max_row or 0) + 1):
        label = _text(hist.cell(rr, idx["fiscal_label"]).value)
        if not _QUARTER_LABEL_RE.fullmatch(label):
            continue
        record = {header: hist.cell(rr, col).value for header, col in idx.items()}
        record["_quarter_key"] = str(hist.cell(rr, idx["quarter"]).value)[:10]
        records[label] = record
    assert records, "History_Q should expose fiscal-label source history"
    return records


def _data_fact_period_keys(wb: Any) -> set[str]:
    ws = wb["DATA_Facts_Long"]
    headers = [_text(ws.cell(1, cc).value) for cc in range(1, int(ws.max_column or 0) + 1)]
    idx = {header: pos + 1 for pos, header in enumerate(headers) if header}
    assert {"period_end", "value"}.issubset(idx), "DATA_Facts_Long should expose period_end and value"
    keys: set[str] = set()
    for rr in range(2, int(ws.max_row or 0) + 1):
        if ws.cell(rr, idx["value"]).value in (None, ""):
            continue
        period_key = str(ws.cell(rr, idx["period_end"]).value)[:10]
        if period_key:
            keys.add(period_key)
    assert keys, "DATA_Facts_Long should expose non-empty source facts"
    return keys


def _anf_period_index_labels(wb: Any) -> list[str]:
    records = _history_records_by_fiscal_label(wb)
    label_by_period_key = {
        _text(record.get("_quarter_key")): label for label, record in records.items() if _text(record.get("_quarter_key"))
    }
    ws = wb["DATA_Period_Index"]
    headers = [_text(ws.cell(1, cc).value) for cc in range(1, int(ws.max_column or 0) + 1)]
    idx = {header: pos + 1 for pos, header in enumerate(headers) if header}
    assert {"period_end", "display_order"}.issubset(idx), "DATA_Period_Index should expose period_end/display_order"
    ordered: list[tuple[int, str]] = []
    for rr in range(2, int(ws.max_row or 0) + 1):
        period_key = str(ws.cell(rr, idx["period_end"]).value)[:10]
        label = label_by_period_key.get(period_key)
        assert label, f"DATA_Period_Index period {period_key!r} should resolve to History_Q fiscal_label"
        ordered.append((int(ws.cell(rr, idx["display_order"]).value or len(ordered) + 1), label))
    labels = [label for _order, label in sorted(ordered)]
    assert labels, "DATA_Period_Index should define the ANF rolling visible window"
    assert labels == sorted(labels, key=_quarter_key), "DATA_Period_Index labels should be ordered"
    return labels


def _anf_visible_windows(wb: Any) -> dict[str, list[str]]:
    period_index = _anf_period_index_labels(wb)

    valuation = wb["Valuation"]
    valuation_labels = _quarter_labels_from_header(valuation, _find_row(valuation, "Quarter"))
    assert valuation_labels == period_index, "ANF Valuation should mirror DATA_Period_Index rolling quarters"

    operating = wb["Operating_Drivers"]
    operating_section = _find_row_contains(operating, "Actuals", start=1)
    operating_labels = _quarter_labels_from_header(operating, _find_row(operating, "Quarter", start=operating_section))
    assert operating_labels == period_index, "ANF Operating_Drivers should mirror DATA_Period_Index rolling quarters"

    bs_segments = wb["BS_Segments"]
    bs_labels = _quarter_labels_from_header(bs_segments, _find_row(bs_segments, "Quarter"))
    assert bs_labels == period_index[-len(bs_labels) :], "ANF BS_Segments should use the shorter trailing window"
    assert 0 < len(bs_labels) < len(period_index), "ANF BS_Segments should remain a shorter rolling window"

    return {
        "period_index": period_index,
        "valuation": valuation_labels,
        "operating_drivers": operating_labels,
        "bs_segments": bs_labels,
    }


def _assert_prior_year_history_and_facts_exist(wb: Any, visible_labels: list[str], sheet_labels: list[str]) -> None:
    history = _history_records_by_fiscal_label(wb)
    fact_period_keys = _data_fact_period_keys(wb)
    for label in visible_labels:
        prior_label = _prior_year_quarter(label)
        assert prior_label not in sheet_labels, f"{prior_label} should be hidden for visible {label}"
        assert prior_label in history, f"History_Q should keep hidden prior-year source history for {label}"
        period_key = _text(history[prior_label].get("_quarter_key"))
        assert period_key in fact_period_keys, f"DATA_Facts_Long should keep source facts for hidden {prior_label}"


def _history_millions(record: dict[str, Any], field: str) -> float:
    value = _num(record.get(field))
    assert value is not None, f"History_Q {field} should be populated"
    return value / 1_000_000.0


def _history_ratio(record: dict[str, Any], numerator: str, denominator: str) -> float:
    numerator_value = _num(record.get(numerator))
    denominator_value = _num(record.get(denominator))
    assert numerator_value is not None, f"History_Q {numerator} should be populated"
    assert denominator_value is not None and abs(denominator_value) > 1e-12, (
        f"History_Q {denominator} should be populated for ratio checks"
    )
    return numerator_value / denominator_value


def _history_ttm_labels(history: dict[str, dict[str, Any]], label: str) -> list[str]:
    labels = sorted(history, key=_quarter_key)
    assert label in history, f"History_Q should include visible label {label}"
    idx = labels.index(label)
    assert idx >= 3, f"History_Q should have four-quarter TTM history for {label}"
    return labels[idx - 3 : idx + 1]


def _find_segment_row_after_group(ws: Worksheet, group: str, segment: str, *, start: int) -> int:
    group_row = _find_row(ws, group, start=start)
    for rr in range(group_row + 1, min(group_row + 10, int(ws.max_row or 0)) + 1):
        if _text(ws.cell(rr, 1).value) == segment:
            return rr
    raise AssertionError(f"{ws.title}: segment {segment!r} missing under {group!r}")


def _fill_rgb(cell: Any) -> str:
    try:
        return str(cell.fill.fgColor.rgb or "").upper()
    except Exception:
        return ""


def _has_bucket_fill(cell: Any) -> bool:
    return str(getattr(cell.fill, "fill_type", "") or "").lower() == "solid" and _fill_rgb(cell) not in {
        "",
        "00000000",
        "00FFFFFF",
        "FFFFFFFF",
    }


def _promise_row(ws: Worksheet, section: str, metric: str) -> int:
    section_row = _find_row(ws, section)
    header_row = _find_row(ws, "Metric", start=section_row)
    for rr in range(header_row + 1, int(ws.max_row or 0) + 1):
        first = _text(ws.cell(rr, 1).value)
        if rr > header_row + 1 and first.endswith("revisions"):
            break
        if first == metric:
            return rr
    raise AssertionError(f"{ws.title}: metric {metric!r} missing under {section!r}")


def test_hidden_source_comparison_uses_ordered_fiscal_history_but_not_sparse_wrong_quarters() -> None:
    fiscal_source = {
        "2022-04-30": 100.0,
        "2022-07-30": 100.0,
        "2022-10-29": 100.0,
        "2023-01-28": 100.0,
        "2023-04-29": 115.0,
    }
    assert _hidden_source_comparison_metric(
        current_key="2023-04-29",
        current_value=115.0,
        visible_idx=0,
        comparison_basis="yoy",
        directionality="higher_better",
        source_values=fiscal_source,
    ) == pytest.approx(0.15)
    calendar_keyed_source = {
        "2022-03-31": 100.0,
        "2022-06-30": 100.0,
        "2022-09-30": 100.0,
        "2022-12-31": 100.0,
        "2023-03-31": 115.0,
    }
    assert _hidden_source_comparison_metric(
        current_key="2023-04-29",
        current_value=115.0,
        visible_idx=0,
        comparison_basis="yoy",
        directionality="higher_better",
        source_values=calendar_keyed_source,
    ) == pytest.approx(0.15)

    sparse_source = {
        "2021-12-31": 90.0,
        "2022-03-31": 95.0,
        "2022-06-30": 100.0,
        "2023-03-31": 110.0,
        "2023-09-30": 120.0,
    }
    assert _hidden_source_comparison_metric(
        current_key="2023-09-30",
        current_value=120.0,
        visible_idx=1,
        comparison_basis="yoy",
        directionality="higher_better",
        source_values=sparse_source,
    ) is None


def test_pbi_operating_drivers_includes_2026_q1_segment_support_from_release() -> None:
    wb = _load_model("PBI")
    ws = wb["Operating_Drivers"]
    section = _find_row_contains(ws, "Segment support", start=1)
    header = _find_row(ws, "Metric / segment", start=section)
    q1_2026 = _quarter_col(ws, header, "2026-Q1")

    checks = [
        ("Revenue ($m)", "SendTech Solutions", 313.947),
        ("Revenue ($m)", "Presort Services", 163.466),
        ("Adj EBIT / operating profit ($m)", "SendTech Solutions", 113.530),
        ("Adj EBIT / operating profit ($m)", "Presort Services", 39.178),
        ("D&A ($m)", "SendTech Solutions", 9.875),
        ("D&A ($m)", "Presort Services", 8.736),
        ("Adj EBITDA ($m)", "SendTech Solutions", 123.405),
        ("Adj EBITDA ($m)", "Presort Services", 47.914),
    ]
    for group, segment, expected in checks:
        row = _find_segment_row_after_group(ws, group, segment, start=header)
        assert _num(ws.cell(row, q1_2026).value) == pytest.approx(expected, abs=0.01), (
            f"PBI Operating_Drivers {group}/{segment} 2026-Q1 should come from the Q1 2026 release"
        )
    wb.close()


def test_pbi_operating_drivers_reportable_segment_totals_stay_populated() -> None:
    wb = _load_model("PBI")
    ws = wb["Operating_Drivers"]
    section = _find_row_contains(ws, "Segment support", start=1)
    header = _find_row(ws, "Metric / segment", start=section)

    expected_totals = {
        "Revenue ($m)": {
            "2024-Q2": 489.745,
            "2024-Q4": 516.121,
            "2025-Q4": 477.625,
            "2026-Q1": 477.413,
        },
        "Adj EBIT / operating profit ($m)": {
            "2024-Q2": 118.950,
            "2024-Q4": 142.386,
            "2025-Q4": 154.780,
            "2026-Q1": 152.708,
        },
        "D&A ($m)": {
            "2024-Q2": 20.524,
            "2024-Q4": 20.534,
            "2025-Q4": 20.303,
            "2026-Q1": 18.611,
        },
        "Adj EBITDA ($m)": {
            "2024-Q2": 139.474,
            "2024-Q4": 162.920,
            "2025-Q4": 175.083,
            "2026-Q1": 171.319,
        },
    }
    for group, quarter_values in expected_totals.items():
        row = _find_segment_row_after_group(ws, group, "Total reportable segments", start=header)
        for quarter, expected in quarter_values.items():
            col = _quarter_col(ws, header, quarter)
            assert _num(ws.cell(row, col).value) == pytest.approx(expected, abs=0.02), (
                f"PBI Operating_Drivers {group}/Total reportable segments {quarter} should stay source-backed"
            )
    wb.close()


def test_anf_valuation_first_visible_quarters_use_hidden_prior_year_bucket_fills() -> None:
    wb = _load_model_with_styles("ANF")
    try:
        windows = _anf_visible_windows(wb)
        ws = wb["Valuation"]
        header = _find_row(ws, "Quarter")
        visible_quarters = windows["valuation"][:4]
        quarter_cols = {q: _quarter_col(ws, header, q) for q in visible_quarters}
        _assert_prior_year_history_and_facts_exist(wb, visible_quarters, windows["valuation"])

        fill_metrics = (
            "Revenue",
            "Gross margin %",
            "Operating margin %",
            "Operating margin (TTM)",
            "EBITDA margin %",
            "Adj EBITDA margin %",
            "EBIT margin %",
            "Net income attrib. to A&F margin %",
            "Capex % of revenue",
            "FCF (CFO-Capex)",
            "Owner earnings (proxy)",
            "FCF margin %",
            "Current ratio",
            "EPS (GAAP)",
            "Adj EPS",
            "BV/share",
            "FCF/share (TTM)",
            "Net leverage",
            "Net leverage (Adj)",
        )
        source_checks = {
            "Revenue": (lambda rec: _history_millions(rec, "revenue"), 0.0005),
            "Gross margin %": (lambda rec: _history_ratio(rec, "gross_profit", "revenue"), 0.0005),
            "Operating margin %": (lambda rec: _history_ratio(rec, "op_income", "revenue"), 0.0005),
            "EBITDA margin %": (lambda rec: _history_ratio(rec, "ebitda", "revenue"), 0.0005),
            "Adj EBITDA margin %": (lambda rec: _history_ratio(rec, "ebitda", "revenue"), 0.0005),
            "EBIT margin %": (lambda rec: _history_ratio(rec, "op_income", "revenue"), 0.0005),
            "Net income attrib. to A&F margin %": (lambda rec: _history_ratio(rec, "net_income", "revenue"), 0.0005),
            "Capex % of revenue": (lambda rec: _history_ratio(rec, "capex", "revenue"), 0.0005),
            "FCF (CFO-Capex)": (
                lambda rec: _history_millions(rec, "cfo") - _history_millions(rec, "capex"),
                0.0005,
            ),
            "FCF margin %": (
                lambda rec: (
                    (_history_millions(rec, "cfo") - _history_millions(rec, "capex"))
                    / _history_millions(rec, "revenue")
                ),
                0.0005,
            ),
            "Current ratio": (lambda rec: _history_ratio(rec, "assets_current", "liabilities_current"), 0.0005),
            "EPS (GAAP)": (lambda rec: _history_ratio(rec, "net_income", "shares_diluted"), 0.01),
        }
        history = _history_records_by_fiscal_label(wb)
        for metric in fill_metrics:
            row = _find_row(ws, metric)
            for quarter in quarter_cols:
                cell = ws.cell(row, quarter_cols[quarter])
                assert _has_bucket_fill(cell), (
                    f"ANF Valuation {metric} {quarter} should use hidden prior-year source history for comparison fill"
                )
        for metric, (expected_from_history, tolerance) in source_checks.items():
            row = _find_row(ws, metric)
            for quarter, col in quarter_cols.items():
                assert _num(ws.cell(row, col).value) == pytest.approx(
                    expected_from_history(history[quarter]), abs=tolerance
                ), f"ANF Valuation {metric} {quarter} should match History_Q source data"
    finally:
        wb.close()


def test_anf_operating_drivers_first_visible_brand_geography_rows_are_colored_from_source_comps() -> None:
    wb = _load_model_with_styles("ANF")
    try:
        windows = _anf_visible_windows(wb)
        ws = wb["Operating_Drivers"]
        section = _find_row_contains(ws, "Actuals", start=1)
        header = _find_row(ws, "Quarter", start=section)
        visible_quarters = windows["operating_drivers"][:4]
        quarter_cols = {q: _quarter_col(ws, header, q) for q in visible_quarters}
        _assert_prior_year_history_and_facts_exist(wb, visible_quarters, windows["operating_drivers"])
        metrics = (
            "Americas sales",
            "EMEA sales",
            "APAC sales",
            "Abercrombie sales",
            "Hollister sales",
            "APAC sales YoY",
            "Americas sales YoY",
            "EMEA sales YoY",
            "Abercrombie sales YoY",
            "Hollister sales YoY",
            "Total comp",
            "Abercrombie comp",
            "Hollister comp",
        )
        for metric in metrics:
            row = _find_row(ws, metric, start=header)
            for quarter, col in quarter_cols.items():
                cell = ws.cell(row, col)
                assert cell.value not in (None, ""), f"ANF Operating_Drivers {metric} {quarter} should stay populated"
                assert _has_bucket_fill(cell), (
                    f"ANF Operating_Drivers {metric} {quarter} should use source-backed YoY/comp evidence for comparison fill"
                )
    finally:
        wb.close()


def test_pbi_operating_drivers_first_visible_segment_quarters_color_only_with_hidden_prior_year_source() -> None:
    wb = _load_model_with_styles("PBI")
    try:
        ws = wb["Operating_Drivers"]
        section = _find_row_contains(ws, "Segment support", start=1)
        header = _find_row(ws, "Metric / segment", start=section)
        quarter_cols = {q: _quarter_col(ws, header, q) for q in ("2023-Q2", "2023-Q3", "2023-Q4")}
        expected_values = {
            "SendTech Solutions": {"2023-Q2": 348.284, "2023-Q3": 345.147, "2023-Q4": 357.386},
            "Presort Services": {"2023-Q2": 143.107, "2023-Q3": 152.451, "2023-Q4": 163.139},
            "Total reportable segments": {"2023-Q2": 500.831, "2023-Q3": 503.033, "2023-Q4": 526.416},
        }
        for segment, expected_by_quarter in expected_values.items():
            row = _find_segment_row_after_group(ws, "Revenue ($m)", segment, start=header)
            for quarter, expected in expected_by_quarter.items():
                cell = ws.cell(row, quarter_cols[quarter])
                if quarter == "2023-Q2":
                    assert _has_bucket_fill(cell), (
                        f"PBI Operating_Drivers Revenue/{segment} {quarter} should use hidden 2022-Q2 source history for comparison fill"
                    )
                else:
                    assert not _has_bucket_fill(cell), (
                        f"PBI Operating_Drivers Revenue/{segment} {quarter} should stay neutral without a clean hidden prior-year source"
                    )
                assert _num(cell.value) == pytest.approx(expected, abs=0.002), (
                    f"PBI Operating_Drivers Revenue/{segment} {quarter} value changed while fixing style"
                )
    finally:
        wb.close()


def test_valuation_debt_rows_use_known_lower_better_comparison_fills() -> None:
    for ticker in ("ANF", "PBI", "GPRE"):
        wb = _load_model_with_styles(ticker)
        try:
            ws = wb["Valuation"]
            header = _find_row(ws, "Quarter")
            visible_quarters = _quarter_labels_from_header(ws, header)
            quarter_cols = {
                q: _quarter_col(ws, header, q)
                for q in ("2024-Q1", "2024-Q2", "2025-Q1", "2025-Q2")
                if q in visible_quarters
            }
            assert quarter_cols, f"{ticker} should expose comparable debt quarters"
            for metric in ("Debt (core borrowings)", "Net debt (core borrowings)"):
                row = _find_row(ws, metric)
                for quarter, col in quarter_cols.items():
                    cell = ws.cell(row, col)
                    if cell.value in (None, ""):
                        continue
                    prior_quarter = quarter.replace(quarter[:4], str(int(quarter[:4]) - 1), 1)
                    prior_col = _quarter_col(ws, header, prior_quarter) if any(
                        ws.cell(header, c).value == prior_quarter for c in range(1, ws.max_column + 1)
                    ) else None
                    if prior_col is not None:
                        prior_value = _num(ws.cell(row, prior_col).value)
                        if prior_value is not None and abs(prior_value) <= 1e-12:
                            assert not _has_bucket_fill(cell), (
                                f"{ticker} Valuation {metric} {quarter} should stay neutral when prior-year debt base is zero"
                            )
                            continue
                    assert _has_bucket_fill(cell), (
                        f"{ticker} Valuation {metric} {quarter} should be colored as lower-better when comparator exists"
                    )
            # First visible quarter should also be colored when hidden source history exists.
            start_quarter = visible_quarters[0]
            first_col = _quarter_col(ws, header, start_quarter)
            for metric in ("Debt (core borrowings)", "Net debt (core borrowings)"):
                cell = ws.cell(_find_row(ws, metric), first_col)
                if cell.value not in (None, ""):
                    assert _has_bucket_fill(cell), (
                        f"{ticker} Valuation {metric} {start_quarter} should use hidden prior-year source when available"
                    )
        finally:
            wb.close()


def test_gpre_operating_driver_volume_rows_color_only_with_source_backed_yoy() -> None:
    wb = _load_model_with_styles("GPRE")
    try:
        ws = wb["Operating_Drivers"]
        section = _find_row_contains(ws, "Actuals", start=1)
        header = _find_row(ws, "Quarter", start=section)
        quarter_cols = {
            q: _quarter_col(ws, header, q)
            for q in ("2023-Q2", "2023-Q3", "2023-Q4", "2024-Q1", "2024-Q2")
        }
        ethanol_row = _find_row(ws, "Ethanol gallons sold (million gallons)", start=header)
        for quarter in ("2023-Q2", "2023-Q3", "2023-Q4", "2024-Q1"):
            cell = ws.cell(ethanol_row, quarter_cols[quarter])
            assert cell.value in (None, ""), "Ethanol gallons sold should stay blank before exact source values begin"
            assert not _has_bucket_fill(cell), "Blank ethanol gallons sold cells should remain neutral"

        for metric in ("Ultra-high protein (k tons)", "Renewable corn oil (million lbs)"):
            row = _find_row(ws, metric, start=header)
            for quarter in ("2023-Q2", "2023-Q3", "2023-Q4", "2024-Q1", "2024-Q2"):
                cell = ws.cell(row, quarter_cols[quarter])
                if cell.value in (None, ""):
                    continue
                assert _has_bucket_fill(cell), (
                    f"GPRE Operating_Drivers {metric} {quarter} should use source-disclosed YoY comparison when available"
                )
    finally:
        wb.close()


_PBI_Q1_REVENUE_SCHEDULE = """
Pitney Bowes Inc.
Business Segment Revenue
(Unaudited; in thousands)
Three Months Ended March 31,
2026 2025 % Change
Sending Technology Solutions $313,947 $315,606 (1%)
Presort Services 163,466 177,814 (8%)
Total revenue $477,413 $493,420 (3%)
"""

_PBI_Q1_EBITDA_SCHEDULE = """
Pitney Bowes Inc.
Adjusted Segment EBIT & EBITDA
(Unaudited; in thousands)
Three Months Ended March 31,
2026 2025 % change
Adjusted Adjusted Adjusted Adjusted Adjusted Adjusted
Segment Segment Segment Segment Segment Segment
EBIT (1) D&A EBITDA EBIT (1) D&A EBITDA EBIT EBITDA
Sending Technology Solutions $113,530 $9,875 $123,405 $97,027 $11,680 $108,707 17% 14%
Presort Services 39,178 8,736 47,914 54,779 9,269 64,048 (28%) (25%)
Total reportable segments $152,708 $18,611 171,319 $151,806 $20,949 172,755 1% (1%)
Reconciliation of Reported Consolidated Results to Adjusted Results
Adjusted EBIT $130,377
Depreciation and amortization $25,641
Adjusted EBITDA $156,018
Corporate expenses (22,331)
"""


def _parse_money_thousands_for_segment_test(value: Any) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    negative = "(" in text and ")" in text
    normalized = re.sub(r"[^0-9.\-]", "", text)
    if not normalized:
        return None
    parsed = float(normalized) * 1_000.0
    return -abs(parsed) if negative else parsed


def _pbi_q1_source_backed_segment_package() -> tuple[dict[str, Any], list[dict[str, Any]]]:
    q_end = dt.date(2026, 3, 31)
    q_ts = pd.Timestamp(q_end)
    source_doc = "fixture:PBI_Q1_2026_earnings_release.pdf#pages=8-9"
    rows = _parse_local_non_gaap_segment_rows_from_text(_PBI_Q1_REVENUE_SCHEDULE, q_end)
    rows += _parse_local_non_gaap_segment_rows_from_text(_PBI_Q1_EBITDA_SCHEDULE, q_end)
    for row in rows:
        row.update(
            {
                "doc": source_doc,
                "page": 8 if row["metric"] == "revenue" else 9,
                "source": "earnings_release",
            }
        )

    metric_names = {
        "revenue": "Revenue",
        "adj_segment_ebit": "Adjusted EBIT",
        "adj_segment_da": "Depreciation & amortization",
        "adj_segment_ebitda": "Adjusted EBITDA",
    }
    metric_ids = {
        "Revenue": "metric:core:revenue@1",
        "Adjusted EBIT": "metric:business-services:adjusted-segment-ebit@1",
        "Depreciation & amortization": "metric:core:depreciation-amortization@1",
        "Adjusted EBITDA": "metric:core:adjusted-ebitda@1",
    }
    metrics: dict[str, dict[str, dict[pd.Timestamp, float]]] = {}
    source_facts: list[SegmentResidualInputFact] = []
    for row in rows:
        metric_name = metric_names[str(row["metric"])]
        segment_name = str(row["segment"])
        metrics.setdefault(metric_name, {}).setdefault(segment_name, {})[q_ts] = float(row["value"])
        source_facts.append(
            segment_residual_input_fact_from_legacy_row(
                company_id="PBI",
                metric_label=metric_name,
                metric_id=metric_ids[metric_name],
                segment_member=segment_name,
                value_millions=float(row["value"]) / 1_000_000.0,
                period_end=q_end,
                period_id="period:pbi:cy2026-q1@1",
                source_doc=source_doc,
                source_type="earnings-release",
                source_locator=f"page:{row['page']}",
                aggregation_role=(
                    "reported_total"
                    if segment_name == "Total reportable segments"
                    else "component"
                ),
            )
        )

    derivations: list[dict[str, Any]] = []
    metrics = _pbi_repair_total_reportable_segment_quarterly_totals_for_bs(
        metrics,
        source_facts=source_facts,
        derivations_out=derivations,
    )
    _pbi_add_corporate_reconciliation_from_release_text(
        metrics,
        _PBI_Q1_EBITDA_SCHEDULE,
        q_ts,
        _parse_money_thousands_for_segment_test,
    )
    return (
        {
            "metrics": metrics,
            "quarters": [q_end],
            "source_doc": source_doc,
            "source_qd": q_end,
            "segment_derivation_ledger": _segment_residual_ledger_payload(source_facts, derivations),
        },
        rows,
    )


def test_pbi_bs_segments_2026_q1_source_backed_residual_and_corporate_values() -> None:
    release_package, source_rows = _pbi_q1_source_backed_segment_package()
    q4_2025 = pd.Timestamp("2025-12-31")
    q1_2026 = pd.Timestamp("2026-03-31")
    q2_2026 = pd.Timestamp("2026-06-30")
    primary = {
        "metrics": {"Revenue": {"Other operations": {q4_2025: 0.0}}},
        "quarters": [q4_2025.date()],
        "source_doc": "fixture:Historical Segment Financials Q4 2025.xlsx",
    }
    merged = _merge_quarterly_segment_packages_per_period(
        primary=primary,
        authoritative_overlay=release_package,
        supplemental_overlay={},
    )

    total_revenue_row = next(
        row
        for row in source_rows
        if row["segment"] == "Total reportable segments" and row["metric"] == "revenue"
    )
    assert total_revenue_row == {
        "quarter": dt.date(2026, 3, 31),
        "segment": "Total reportable segments",
        "metric": "revenue",
        "value": 477_413_000.0,
        "unit": "USD",
        "period_type": "quarter",
        "source_period_label": "quarter",
        "is_table_total": True,
        "evidence_role": "independent_table_total",
        "doc": "fixture:PBI_Q1_2026_earnings_release.pdf#pages=8-9",
        "page": 8,
        "source": "earnings_release",
    }
    assert 313_947_000.0 + 163_466_000.0 == 477_413_000.0
    assert "PBI_Q1_2026_earnings_release.pdf#pages=8-9" in merged["source_doc"]
    ledger = merged["segment_derivation_ledger"]
    assert ledger["contract_id"] == SEGMENT_RESIDUAL_LEDGER_CONTRACT_ID
    source_facts_by_id = {row["record_id"]: row for row in ledger["source_facts"]}
    revenue_derivation = next(
        row
        for row in ledger["derivations"]
        if row["metric_label"] == "Revenue" and row["target_member"] == "Other operations"
    )
    assert revenue_derivation["rule_id"] == SEGMENT_EXACT_RESIDUAL_RULE_ID
    assert revenue_derivation["classification"] == SEGMENT_EXACT_ZERO_CLASSIFICATION
    assert revenue_derivation["value"] == {"kind": "exact", "value": "0"}
    assert revenue_derivation["period_id"] == "period:pbi:cy2026-q1@1"
    assert revenue_derivation["basis_id"] == "basis:core:reported@1"
    assert revenue_derivation["unit_id"] == "unit:core:usd-millions@1"
    assert revenue_derivation["currency"] == "USD"
    assert revenue_derivation["scope"] == "reportable_segments"
    assert revenue_derivation["direct_total_input_id"] in source_facts_by_id
    assert set(revenue_derivation["direct_component_input_ids"]) <= set(source_facts_by_id)
    assert len(revenue_derivation["direct_component_input_ids"]) == 2
    total_input = source_facts_by_id[revenue_derivation["direct_total_input_id"]]
    component_inputs = [source_facts_by_id[record_id] for record_id in revenue_derivation["direct_component_input_ids"]]
    assert total_input["aggregation_role"] == "reported_total"
    assert total_input["assertion_mode"] == "reported"
    assert total_input["value"] == {"kind": "exact", "value": "477.413"}
    assert {row["value"]["value"] for row in component_inputs} == {"313.947", "163.466"}
    assert all(row["assertion_mode"] == "reported" for row in component_inputs)
    assert total_input["source_document_id"] in revenue_derivation["source_document_ids"]
    assert revenue_derivation["evidence_occurrence_ids"]
    assert all(record_id in source_facts_by_id for record_id in revenue_derivation["input_record_ids"])
    assert "cell" not in revenue_derivation["economic_identity"].lower()
    assert "coordinate" not in revenue_derivation["economic_identity"].lower()

    expected_millions = {
        ("Revenue", "Other operations"): 0.0,
        ("Adjusted EBIT", "Other operations"): 0.0,
        ("Adjusted EBIT", "Corporate expense"): -22.331,
        ("Depreciation & amortization", "Other operations"): 0.0,
        ("Depreciation & amortization", "Corporate expense"): 7.030,
        ("Adjusted EBITDA", "Other operations"): 0.0,
        ("Adjusted EBITDA", "Corporate expense"): -15.301,
    }
    for (metric, segment), expected in expected_millions.items():
        series = merged["metrics"][metric][segment]
        assert series[q1_2026] / 1_000_000.0 == pytest.approx(expected, abs=0.000001)
        assert q2_2026 not in series


def test_pbi_segment_residual_missing_total_stays_missing_and_is_not_rebuilt() -> None:
    q_ts = pd.Timestamp("2026-03-31")
    repaired = _pbi_repair_total_reportable_segment_quarterly_totals_for_bs(
        {
            "Revenue": {
                "SendTech Solutions": {q_ts: 313_947_000.0},
                "Presort Services": {q_ts: 163_466_000.0},
            }
        },
        source_facts=(),
        derivations_out=[],
    )

    assert "Other operations" not in repaired["Revenue"]
    assert "Total reportable segments" not in repaired["Revenue"]


def test_pbi_segment_residual_non_exact_or_incompatible_total_fails_closed() -> None:
    q1 = pd.Timestamp("2026-03-31")
    q2 = pd.Timestamp("2026-06-30")
    non_exact = _pbi_repair_total_reportable_segment_quarterly_totals_for_bs(
        {
            "Revenue": {
                "SendTech Solutions": {q1: 313_947_000.0},
                "Presort Services": {q1: 163_466_000.0},
                "Total reportable segments": {q1: 477_414_000.0},
            }
        },
        source_facts=(),
        derivations_out=[],
    )
    incompatible_period = _pbi_repair_total_reportable_segment_quarterly_totals_for_bs(
        {
            "Revenue": {
                "SendTech Solutions": {q1: 313_947_000.0},
                "Presort Services": {q2: 163_466_000.0},
                "Total reportable segments": {q1: 477_413_000.0},
            }
        },
        source_facts=(),
        derivations_out=[],
    )

    assert "Other operations" not in non_exact["Revenue"]
    assert non_exact["Revenue"]["Total reportable segments"][q1] == 477_414_000.0
    assert "Other operations" not in incompatible_period["Revenue"]


def test_segment_residual_typed_contract_rejects_incompatible_or_circular_inputs() -> None:
    common = {
        "company_id": "PBI",
        "metric_label": "Revenue",
        "metric_id": "metric:core:revenue@1",
        "period_end": dt.date(2026, 3, 31),
        "period_id": "period:pbi:cy2026-q1@1",
        "source_doc": "fixture:PBI_Q1_2026_earnings_release.pdf",
        "source_type": "earnings-release",
        "source_locator": "page:8",
    }
    total = segment_residual_input_fact_from_legacy_row(
        **common,
        segment_member="Total reportable segments",
        value_millions="477.413",
        aggregation_role="reported_total",
    )
    sendtech = segment_residual_input_fact_from_legacy_row(
        **common,
        segment_member="SendTech Solutions",
        value_millions="313.947",
        aggregation_role="component",
    )
    presort = segment_residual_input_fact_from_legacy_row(
        **common,
        segment_member="Presort Services",
        value_millions="163.466",
        aggregation_role="component",
    )

    forward = derive_exact_zero_segment_residual(
        total=total,
        components=(sendtech, presort),
        target_member="Other operations",
    )
    reverse = derive_exact_zero_segment_residual(
        total=total,
        components=(presort, sendtech),
        target_member="Other operations",
    )
    assert forward is not None and reverse is not None
    assert forward.to_dict() == reverse.to_dict()
    assert forward.value == "0"

    incompatible_mutations = (
        replace(presort, period_id="period:pbi:cy2026-q2@1", period_end="2026-06-30"),
        replace(presort, basis_id="basis:core:guided@1"),
        replace(presort, unit_id="unit:core:currency-million@1"),
        replace(presort, currency="EUR"),
        replace(presort, scope="consolidated_company"),
    )
    for incompatible in incompatible_mutations:
        with pytest.raises(SegmentNormalizationError):
            derive_exact_zero_segment_residual(
                total=total,
                components=(sendtech, incompatible),
                target_member="Other operations",
            )

    with pytest.raises(SegmentNormalizationError):
        derive_exact_zero_segment_residual(
            total=total,
            components=(sendtech,),
            target_member="Other operations",
        )
    with pytest.raises(SegmentNormalizationError):
        replace(total, assertion_mode="derived")


def test_pbi_segment_exact_residual_and_package_merge_are_source_order_independent() -> None:
    q_ts = pd.Timestamp("2026-03-31")
    ordered_segments = [
        ("SendTech Solutions", 313_947_000.0),
        ("Presort Services", 163_466_000.0),
        ("Total reportable segments", 477_413_000.0),
    ]

    def _package(rows: list[tuple[str, float]]) -> dict[str, Any]:
        source_facts = [
            segment_residual_input_fact_from_legacy_row(
                company_id="PBI",
                metric_label="Revenue",
                metric_id="metric:core:revenue@1",
                segment_member=segment,
                value_millions=value / 1_000_000.0,
                period_end=q_ts.date(),
                period_id="period:pbi:cy2026-q1@1",
                source_doc="fixture:source-order-independent-release",
                source_type="earnings-release",
                source_locator=f"table:revenue:{segment}",
                aggregation_role=(
                    "reported_total"
                    if segment == "Total reportable segments"
                    else "component"
                ),
            )
            for segment, value in rows
        ]
        derivations: list[dict[str, Any]] = []
        repaired = _pbi_repair_total_reportable_segment_quarterly_totals_for_bs(
            {"Revenue": {segment: {q_ts: value} for segment, value in rows}},
            source_facts=source_facts,
            derivations_out=derivations,
        )
        return {
            "metrics": repaired,
            "quarters": [q_ts.date()],
            "source_doc": "fixture:source-order-independent-release",
            "segment_derivation_ledger": _segment_residual_ledger_payload(
                source_facts,
                derivations,
            ),
        }

    forward = _merge_quarterly_segment_packages_per_period(
        primary={},
        authoritative_overlay=_package(ordered_segments),
        supplemental_overlay={},
    )
    reverse = _merge_quarterly_segment_packages_per_period(
        primary={},
        authoritative_overlay=_package(list(reversed(ordered_segments))),
        supplemental_overlay={},
    )

    assert forward == reverse
    assert forward["metrics"]["Revenue"]["Other operations"][q_ts] == 0.0


def test_pbi_unlineaged_primary_cannot_mint_or_suppress_lineaged_exact_zero() -> None:
    q_ts = pd.Timestamp("2026-03-31")
    raw_primary_metrics = {
        "Revenue": {
            "SendTech Solutions": {q_ts: 313_947_000.0},
            "Presort Services": {q_ts: 163_466_000.0},
            "Total reportable segments": {q_ts: 477_413_000.0},
        }
    }
    unlineaged_derivations: list[dict[str, Any]] = []
    fail_closed = _pbi_repair_total_reportable_segment_quarterly_totals_for_bs(
        raw_primary_metrics,
        source_facts=(),
        derivations_out=unlineaged_derivations,
    )
    assert "Other operations" not in fail_closed["Revenue"]
    assert unlineaged_derivations == []

    lineaged_overlay, _ = _pbi_q1_source_backed_segment_package()
    stale_unlineaged_primary = {
        "metrics": {
            "Revenue": {
                **raw_primary_metrics["Revenue"],
                "Other operations": {q_ts: 0.0},
            }
        },
        "quarters": [q_ts.date()],
        "source_doc": "fixture:unlineaged-primary",
    }
    merged = _merge_quarterly_segment_packages_per_period(
        primary=stale_unlineaged_primary,
        authoritative_overlay=lineaged_overlay,
        supplemental_overlay={},
    )

    assert merged["metrics"]["Revenue"]["Other operations"][q_ts] == 0.0
    revenue_derivations = [
        row
        for row in merged["segment_derivation_ledger"]["derivations"]
        if row["metric_label"] == "Revenue"
        and row["target_member"] == "Other operations"
        and row["period_end"] == "2026-03-31"
    ]
    assert len(revenue_derivations) == 1
    assert revenue_derivations[0]["classification"] == SEGMENT_EXACT_ZERO_CLASSIFICATION


def test_pbi_merge_rejects_unresolved_duplicate_or_nonreplaying_residual_ledgers() -> None:
    q_ts = pd.Timestamp("2026-03-31")
    valid_overlay, _ = _pbi_q1_source_backed_segment_package()
    valid_ledger = deepcopy(valid_overlay["segment_derivation_ledger"])
    assert validate_segment_residual_ledger_payload(valid_ledger) == valid_ledger

    revenue_derivation = next(
        row
        for row in valid_ledger["derivations"]
        if row["metric_label"] == "Revenue" and row["target_member"] == "Other operations"
    )
    missing_total = deepcopy(valid_ledger)
    missing_total["source_facts"] = [
        row
        for row in missing_total["source_facts"]
        if row["record_id"] != revenue_derivation["direct_total_input_id"]
    ]

    missing_component = deepcopy(valid_ledger)
    missing_component["source_facts"] = [
        row
        for row in missing_component["source_facts"]
        if row["record_id"] != revenue_derivation["direct_component_input_ids"][0]
    ]

    duplicate_fact = deepcopy(valid_ledger)
    duplicate_fact["source_facts"].append(deepcopy(duplicate_fact["source_facts"][0]))

    duplicate_semantic_fact = deepcopy(valid_ledger)
    duplicated = deepcopy(duplicate_semantic_fact["source_facts"][0])
    duplicated["record_id"] = f"{duplicated['record_id']}|forged-duplicate"
    duplicate_semantic_fact["source_facts"].append(duplicated)

    nonreplaying_arithmetic = deepcopy(valid_ledger)
    revenue_component = next(
        row
        for row in nonreplaying_arithmetic["source_facts"]
        if row["metric_label"] == "Revenue" and row["segment_member"] == "SendTech Solutions"
    )
    revenue_component["value"]["value"] = "313.946"

    inconsistent_target = deepcopy(valid_ledger)
    inconsistent_target["derivations"][0]["target_member"] = "Unexpected operations"

    unlineaged_primary = {
        "metrics": {"Revenue": {"Other operations": {q_ts: 0.0}}},
        "quarters": [q_ts.date()],
        "source_doc": "fixture:unlineaged-primary",
    }
    for malformed_ledger in (
        missing_total,
        missing_component,
        duplicate_fact,
        duplicate_semantic_fact,
        nonreplaying_arithmetic,
        inconsistent_target,
    ):
        with pytest.raises(SegmentNormalizationError):
            validate_segment_residual_ledger_payload(malformed_ledger)
        malformed_overlay = deepcopy(valid_overlay)
        malformed_overlay["segment_derivation_ledger"] = malformed_ledger
        merged = _merge_quarterly_segment_packages_per_period(
            primary=unlineaged_primary,
            authoritative_overlay=malformed_overlay,
            supplemental_overlay={},
        )
        assert merged["metrics"]["Revenue"]["Other operations"][q_ts] == 0.0
        assert "segment_derivation_ledger" not in merged
        overlay_only = _merge_quarterly_segment_packages_per_period(
            primary={},
            authoritative_overlay=malformed_overlay,
            supplemental_overlay={},
        )
        assert overlay_only == {}

    package_mismatch = deepcopy(valid_overlay)
    package_mismatch["metrics"]["Revenue"]["SendTech Solutions"][q_ts] = 313_946_000.0
    assert validate_segment_residual_ledger_payload(
        package_mismatch["segment_derivation_ledger"]
    ) == valid_ledger
    merged_mismatch = _merge_quarterly_segment_packages_per_period(
        primary=unlineaged_primary,
        authoritative_overlay=package_mismatch,
        supplemental_overlay={},
    )
    assert "segment_derivation_ledger" not in merged_mismatch
    overlay_only_mismatch = _merge_quarterly_segment_packages_per_period(
        primary={},
        authoritative_overlay=package_mismatch,
        supplemental_overlay={},
    )
    assert overlay_only_mismatch == {}

    malformed_shape = deepcopy(valid_overlay)
    malformed_shape["segment_derivation_ledger"]["derivations"] = {
        "not": "a deterministic derivation array"
    }
    assert _merge_quarterly_segment_packages_per_period(
        primary={},
        authoritative_overlay=malformed_shape,
        supplemental_overlay={},
    ) == {}


@pytest.mark.parametrize(
    "malformed_ledger",
    (
        "not-a-ledger-mapping",
        ["not", "a", "mapping"],
        {
            "contract_id": SEGMENT_RESIDUAL_LEDGER_CONTRACT_ID,
            "source_facts": "not-an-array",
            "derivations": [],
        },
        {
            "contract_id": SEGMENT_RESIDUAL_LEDGER_CONTRACT_ID,
            "source_facts": [],
            "derivations": "not-an-array",
        },
        None,
    ),
    ids=(
        "scalar",
        "array",
        "malformed-top-level-rows",
        "malformed-derivations",
        "explicit-null",
    ),
)
def test_pbi_overlay_only_malformed_ledger_shape_cannot_own_numeric_residual(
    malformed_ledger: Any,
) -> None:
    q_ts = pd.Timestamp("2026-03-31")
    overlay = {
        "metrics": {"Revenue": {"Other operations": {q_ts: 0.0}}},
        "quarters": [q_ts.date()],
        "source_doc": "fixture:malformed-ledger",
        "segment_derivation_ledger": malformed_ledger,
    }

    with pytest.raises(SegmentNormalizationError):
        validate_segment_residual_ledger_payload(malformed_ledger)
    assert _merge_quarterly_segment_packages_per_period(
        primary={},
        authoritative_overlay=overlay,
        supplemental_overlay={},
    ) == {}


def test_segment_package_ledger_merge_states_do_not_conflate_empty_and_unparseable() -> None:
    valid_package, _ = _pbi_q1_source_backed_segment_package()
    absent = _segment_package_ledger_merge_validation(
        {"metrics": {"Revenue": {"Direct member": {pd.Timestamp("2026-03-31"): 1.0}}}}
    )
    valid = _segment_package_ledger_merge_validation(valid_package)

    enumerable_invalid_package = deepcopy(valid_package)
    enumerable_invalid_package["segment_derivation_ledger"]["source_facts"].append(
        deepcopy(enumerable_invalid_package["segment_derivation_ledger"]["source_facts"][0])
    )
    enumerable_invalid = _segment_package_ledger_merge_validation(enumerable_invalid_package)

    unenumerable_invalid_package = deepcopy(valid_package)
    unenumerable_invalid_package["segment_derivation_ledger"]["source_facts"] = "not-an-array"
    unenumerable_invalid_package["segment_derivation_ledger"]["derivations"] = []
    unenumerable_invalid = _segment_package_ledger_merge_validation(unenumerable_invalid_package)

    malformed_derivations_package = deepcopy(valid_package)
    malformed_derivations_package["segment_derivation_ledger"]["source_facts"] = []
    malformed_derivations_package["segment_derivation_ledger"]["derivations"] = "not-an-array"
    malformed_derivations = _segment_package_ledger_merge_validation(
        malformed_derivations_package
    )

    assert not absent.ledger_present
    assert absent.target_enumeration_valid
    assert absent.package_merge_allowed

    assert valid.ledger_present and valid.ledger_valid
    assert valid.target_enumeration_valid
    assert valid.validated_targets
    assert valid.package_merge_allowed

    assert enumerable_invalid.ledger_present and not enumerable_invalid.ledger_valid
    assert enumerable_invalid.target_enumeration_valid
    assert enumerable_invalid.declared_targets
    assert not enumerable_invalid.package_merge_allowed

    for decision, reason in (
        (unenumerable_invalid, "ledger_source_facts_type_invalid"),
        (malformed_derivations, "ledger_derivations_type_invalid"),
    ):
        assert decision.ledger_present and not decision.ledger_valid
        assert not decision.target_enumeration_valid
        assert decision.declared_targets == frozenset()
        assert decision.invalid_reason == reason
        assert not decision.package_merge_allowed


def test_malformed_package_precedence_is_atomic_and_valid_overlay_remains_authoritative() -> None:
    q_ts = pd.Timestamp("2026-03-31")
    valid_package, _ = _pbi_q1_source_backed_segment_package()

    malformed = deepcopy(valid_package)
    malformed["segment_derivation_ledger"]["source_facts"] = "not-an-array"
    malformed["segment_derivation_ledger"]["derivations"] = []
    malformed["metrics"]["Revenue"]["Unrelated direct-looking member"] = {q_ts: 25.0}

    valid_primary = _merge_quarterly_segment_packages_per_period(
        primary=valid_package,
        authoritative_overlay=malformed,
        supplemental_overlay={},
    )
    assert valid_primary["metrics"]["Revenue"]["Other operations"][q_ts] == 0.0
    assert "Unrelated direct-looking member" not in valid_primary["metrics"]["Revenue"]
    assert valid_primary["segment_derivation_ledger"] == valid_package["segment_derivation_ledger"]

    valid_overlay = _merge_quarterly_segment_packages_per_period(
        primary=malformed,
        authoritative_overlay=valid_package,
        supplemental_overlay={},
    )
    assert valid_overlay["metrics"]["Revenue"]["Other operations"][q_ts] == 0.0
    assert "Unrelated direct-looking member" not in valid_overlay["metrics"]["Revenue"]
    assert valid_overlay["segment_derivation_ledger"] == valid_package["segment_derivation_ledger"]


def test_ledger_absent_direct_package_still_merges_but_invalid_claiming_package_is_atomic() -> None:
    q_ts = pd.Timestamp("2026-03-31")
    direct = {
        "metrics": {"Revenue": {"Independent direct member": {q_ts: 25.0}}},
        "quarters": [q_ts.date()],
        "source_doc": "fixture:independent-direct-package",
    }
    assert _merge_quarterly_segment_packages_per_period(
        primary=direct,
        authoritative_overlay={},
        supplemental_overlay={},
    )["metrics"]["Revenue"]["Independent direct member"][q_ts] == 25.0

    invalid_claim = deepcopy(direct)
    invalid_claim["segment_derivation_ledger"] = {
        "contract_id": SEGMENT_RESIDUAL_LEDGER_CONTRACT_ID,
        "source_facts": "not-an-array",
        "derivations": [],
    }
    assert _merge_quarterly_segment_packages_per_period(
        primary=invalid_claim,
        authoritative_overlay={},
        supplemental_overlay={},
    ) == {}


def test_pbi_residual_contract_does_not_mint_other_for_anf_or_gpre_segments() -> None:
    q_ts = pd.Timestamp("2026-03-31")
    for segments in (
        {"Americas": {q_ts: 600.0}, "EMEA": {q_ts: 200.0}, "APAC": {q_ts: 100.0}},
        {"Ethanol production": {q_ts: 300.0}, "Agribusiness": {q_ts: 50.0}},
    ):
        metrics = {"Revenue": segments}
        assert _pbi_repair_total_reportable_segment_quarterly_totals_for_bs(
            metrics,
            source_facts=(),
            derivations_out=[],
        ) == metrics


def test_pbi_promise_adjusted_ebit_and_eps_progress_are_source_backed() -> None:
    wb = _load_model("PBI")
    ws = wb["Promise_Progress_UI"]

    expected_rows = {
        ("2025-Q2 revisions", "Adjusted EBIT guidance"): ("$102.3m", "YTD: $222.0m"),
        ("2025-Q3 revisions", "Adjusted EBIT guidance"): ("$107.3m", "YTD: $329.3m"),
        ("2024-Q3 revisions", "Adjusted EBIT guidance"): ("$102.8m", "YTD: $270.8m"),
        ("2026-Q1 revisions", "Adjusted EPS guidance"): ("$0.47", "YTD: $0.47"),
    }
    for (section, metric), (actual, progress) in expected_rows.items():
        row = _promise_row(ws, section, metric)
        assert _text(ws.cell(row, 5).value) == actual
        assert _text(ws.cell(row, 6).value) == progress
        assert _text(ws.cell(row, 7).value) != "Completed"
    wb.close()


def test_pbi_promise_q4_rows_split_quarter_actual_from_fy_progress() -> None:
    wb = _load_model("PBI")
    ws = wb["Promise_Progress_UI"]

    expected_final_2025 = {
        "Revenue guidance": ("$478m", "FY: $1.89bn", "Missed"),
        "Adjusted EBIT guidance": ("$132m", "FY: $461.3m", "Hit"),
        "Adjusted EPS guidance": ("$0.45", "FY: $1.35", "Hit"),
        "FCF target": ("$212m", "FY: $358.3m", "Hit"),
    }
    for metric, (actual, progress, status) in expected_final_2025.items():
        row = _promise_row(ws, "2025-Q4 revisions", metric)
        assert _text(ws.cell(row, 4).value) == "Completed"
        assert _text(ws.cell(row, 5).value) == actual
        assert _text(ws.cell(row, 6).value) == progress
        assert _text(ws.cell(row, 7).value) == status
        assert _text(ws.cell(row, 8).value) == "2025 year"
        assert "Q4 actual shown in Actual" in _text(ws.cell(row, 11).value)

    row = _promise_row(ws, "2024-Q4 revisions", "Adjusted EBIT guidance")
    assert _text(ws.cell(row, 4).value) == "Completed"
    assert _text(ws.cell(row, 5).value) == "$114m"
    assert _text(ws.cell(row, 6).value) == "FY: $385.2m"
    assert _text(ws.cell(row, 7).value) == "Beat"
    wb.close()


def test_anf_operating_drivers_keeps_brand_and_geography_q4_values() -> None:
    wb = _load_model("ANF")
    ws = wb["Operating_Drivers"]
    section = _find_row_contains(ws, "Actuals", start=1)
    header = _find_row(ws, "Quarter", start=section)

    expected_by_row = {
        "Americas sales": {"2023-Q4": 1193.3, "2024-Q4": 1319.7, "2025-Q4": 1383.9},
        "Abercrombie sales": {"2023-Q4": 755.2, "2024-Q4": 772.7, "2025-Q4": 806.5},
        "Hollister sales": {"2023-Q4": 697.7, "2024-Q4": 812.2, "2025-Q4": 863.3},
        "Net sales": {"2023-Q4": 1452.9, "2024-Q4": 1584.9, "2025-Q4": 1669.8},
    }
    for row_label, quarter_values in expected_by_row.items():
        row = _find_row(ws, row_label, start=header)
        for quarter, expected in quarter_values.items():
            col = _quarter_col(ws, header, quarter)
            assert _num(ws.cell(row, col).value) == pytest.approx(expected, abs=0.1), (
                f"ANF Operating_Drivers {row_label} {quarter} should not lose source-backed Q4 data"
            )

    for quarter in ("2023-Q4", "2024-Q4", "2025-Q4"):
        col = _quarter_col(ws, header, quarter)
        abercrombie = _num(ws.cell(_find_row(ws, "Abercrombie sales", start=header), col).value)
        hollister = _num(ws.cell(_find_row(ws, "Hollister sales", start=header), col).value)
        net_sales = _num(ws.cell(_find_row(ws, "Net sales", start=header), col).value)
        assert abercrombie is not None and hollister is not None and net_sales is not None
        assert abercrombie + hollister == pytest.approx(net_sales, abs=0.3)
    wb.close()


def test_anf_promise_q4_splits_quarter_actual_from_fy_progress() -> None:
    wb = _load_model("ANF")
    ws = wb["Promise_Progress_UI"]

    expected_final = {
        "Net sales growth": ("+5.4%", "FY: +6%", "Completed"),
        "Operating margin": ("14.1%", "FY: 13.3% GAAP / 12.5% adjusted", "Mixed"),
        "Adjusted EPS": ("$3.68 adjusted", "FY: $9.86 adjusted", "Missed"),
        "Capex": ("$55.6m", "FY: $240.8m", "Hit"),
        "Diluted shares": ("46.8m diluted", "Δ vs guide: -1.2m; Δ YTD: -5.6m", "Completed"),
        "Share repurchases": ("$100.0m", "FY: $450m", "Completed"),
    }
    for metric, (actual, progress, status) in expected_final.items():
        row = _promise_row(ws, "2025-Q4 revisions", metric)
        assert _text(ws.cell(row, 5).value) == actual
        assert _text(ws.cell(row, 6).value) == progress
        assert _text(ws.cell(row, 7).value) == status
        assert _text(ws.cell(row, 8).value) == "2025 year"
        assert _text(ws.cell(row, 10).value) == "2026-03-04"
        assert "Q4 actual shown in Actual" in _text(ws.cell(row, 11).value)

    expected_pre_release = {
        "Net sales growth": ("+5.4%", "FY: +6%"),
        "Operating margin": ("14.1%", "FY: 13.3% GAAP / 12.5% adjusted"),
        "Adjusted EPS": ("$3.68 adjusted", "FY: $9.86 adjusted"),
        "Capex": ("$55.6m", "FY: $240.8m"),
    }
    for metric, (actual, progress) in expected_pre_release.items():
        row = _promise_row(ws, "2025-Q4 pre-release update revisions", metric)
        assert _text(ws.cell(row, 5).value) == actual
        assert _text(ws.cell(row, 6).value) == progress
        assert _text(ws.cell(row, 7).value) == "On track"
        assert _text(ws.cell(row, 10).value) == "2026-01-12"
        assert "pre-release was issued before final report" in _text(ws.cell(row, 11).value)

    for section, metric, expected_actual, expected_progress in (
        ("2025-Q1 revisions", "Adjusted EPS", "$1.59 adjusted", "YTD: $1.59 adjusted"),
        ("2025-Q2 revisions", "Adjusted EPS", "$2.32 adjusted", "YTD: $3.91 adjusted"),
        ("2025-Q3 revisions", "Adjusted EPS", "$2.36 adjusted", "YTD: $6.27 adjusted"),
    ):
        row = _promise_row(ws, section, metric)
        assert _text(ws.cell(row, 5).value) == expected_actual
        assert _text(ws.cell(row, 6).value) == expected_progress
        assert _text(ws.cell(row, 7).value) == "On track"

    for section, expected_actual, expected_progress in (
        ("2025-Q1 revisions", "50.6m diluted", "Δ vs guide: +1.6m; Δ YTD: -1.8m"),
        ("2025-Q2 revisions", "48.6m diluted", "Δ vs guide: -0.4m; Δ YTD: -3.9m"),
        ("2025-Q3 revisions", "47.9m diluted", "Δ vs guide: -0.1m; Δ YTD: -4.6m"),
    ):
        row = _promise_row(ws, section, "Diluted shares")
        assert _text(ws.cell(row, 5).value) == expected_actual
        assert _text(ws.cell(row, 6).value) == expected_progress
        assert _text(ws.cell(row, 7).value) == "On track"
    wb.close()


def test_anf_promise_timeline_rows_keep_hidden_source_keys_aligned() -> None:
    wb = _load_model("ANF")
    ws = wb["Promise_Progress_UI"]
    rows = []
    current_section = ""
    for rr in range(1, int(ws.max_row or 0) + 1):
        first = _text(ws.cell(rr, 1).value)
        if first.endswith("revisions"):
            current_section = first
            continue
        if not current_section or first in {"", "Metric"}:
            continue
        source_date = _text(ws.cell(rr, 10).value)
        source_note = _text(ws.cell(rr, 11).value)
        if not source_date and not source_note:
            continue
        rows.append((rr, current_section, first))

    assert rows, "ANF Promise_Progress_UI should have source-backed timeline rows"
    blank_keys = [f"{section}!A{rr} {metric}" for rr, section, metric in rows if not _text(ws.cell(rr, 15).value)]
    assert not blank_keys, "ANF source-backed timeline rows should retain hidden source keys: " + ", ".join(blank_keys[:8])

    bad_visible_key_values = {"on track", "completed", "hit", "missed", "mixed", "open", "2025 year"}
    for rr, _section, _metric in rows:
        hidden_key = _text(ws.cell(rr, 15).value)
        assert hidden_key.startswith("guidance:"), f"ANF Promise_Progress_UI!O{rr} has misaligned hidden key {hidden_key!r}"
        assert hidden_key.lower() not in bad_visible_key_values
        assert not re.fullmatch(r"20\d{2}-\d{2}-\d{2}", hidden_key)
        assert not hidden_key.startswith(("$", "+"))
        metric_slug = re.sub(r"[^a-z0-9]+", "_", _metric.lower()).strip("_")
        source_date_slug = re.sub(r"[^a-z0-9]+", "_", _text(ws.cell(rr, 10).value).lower()).strip("_")
        assert f":{metric_slug}:" in hidden_key, (
            f"ANF Promise_Progress_UI!O{rr} key {hidden_key!r} should align to visible metric {_metric!r}"
        )
        if source_date_slug:
            assert hidden_key.endswith(source_date_slug), (
                f"ANF Promise_Progress_UI!O{rr} key {hidden_key!r} should align to source date {source_date_slug!r}"
            )

    q4_expectations = {
        ("2025-Q4 revisions", "Net sales growth"): ("+5.4%", "FY: +6%", "Completed"),
        ("2025-Q4 revisions", "Capex"): ("$55.6m", "FY: $240.8m", "Hit"),
        ("2025-Q4 revisions", "Share repurchases"): ("$100.0m", "FY: $450m", "Completed"),
        ("2025-Q4 pre-release update revisions", "Adjusted EPS"): ("$3.68 adjusted", "FY: $9.86 adjusted", "On track"),
    }
    for (section, metric), (actual, progress, status) in q4_expectations.items():
        row = _promise_row(ws, section, metric)
        assert _text(ws.cell(row, 5).value) == actual
        assert _text(ws.cell(row, 6).value) == progress
        assert _text(ws.cell(row, 7).value) == status
        assert _text(ws.cell(row, 15).value)
    wb.close()


def test_anf_valuation_adjusted_eps_and_ttm_uses_source_backed_older_quarters() -> None:
    wb = _load_model("ANF")
    ws = wb["Valuation"]
    header = _find_row(ws, "Quarter")
    adj_eps_row = _find_row(ws, "Adj EPS", start=header)
    ebitda_ttm_row = _find_row(ws, "EBITDA (TTM)", start=header)
    adj_ebitda_ttm_row = _find_row(ws, "Adj EBITDA (TTM)", start=header)

    windows = _anf_visible_windows(wb)
    visible_quarters = windows["valuation"]
    first_visible = visible_quarters[0]
    first_col = _quarter_col(ws, header, first_visible)
    assert _num(ws.cell(adj_eps_row, first_col).value) is not None, (
        f"ANF Valuation Adj EPS {first_visible} should stay populated"
    )

    history = _history_records_by_fiscal_label(wb)
    ttm_labels = _history_ttm_labels(history, first_visible)
    hidden_ttm_labels = [label for label in ttm_labels if label not in visible_quarters]
    assert hidden_ttm_labels, f"ANF {first_visible} TTM should use hidden older History_Q quarters"
    expected_ebitda_ttm = sum(_history_millions(history[label], "ebitda") for label in ttm_labels)
    assert _num(ws.cell(ebitda_ttm_row, first_col).value) == pytest.approx(expected_ebitda_ttm, abs=0.002)
    assert _num(ws.cell(adj_ebitda_ttm_row, first_col).value) == pytest.approx(expected_ebitda_ttm, abs=0.002)

    hist = wb["History_Q"]
    headers = [_text(hist.cell(1, cc).value) for cc in range(1, int(hist.max_column or 0) + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    assert {"quarter", "fiscal_label", "da", "ebitda"}.issubset(idx), (
        "History_Q should expose D&A, EBITDA, and fiscal labels"
    )
    hidden_source_row = None
    hidden_source_label = hidden_ttm_labels[0]
    for rr in range(2, int(hist.max_row or 0) + 1):
        if _text(hist.cell(rr, idx["fiscal_label"]).value) == hidden_source_label:
            hidden_source_row = rr
            break
    assert hidden_source_row is not None, f"ANF hidden history should keep {hidden_source_label} for TTM support"
    assert _num(hist.cell(hidden_source_row, idx["da"]).value) is not None
    assert _num(hist.cell(hidden_source_row, idx["ebitda"]).value) is not None
    wb.close()


def test_gpre_visible_adjusted_metrics_do_not_show_micro_transcript_artifacts() -> None:
    wb = _load_model("GPRE")
    for sheet_name in ("Valuation", "Operating_Drivers"):
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            row_label = _text(row[0].value).lower() if row else ""
            if not row_label or not any(token in row_label for token in ("adj ebitda", "adj fcf", "adj eps")):
                continue
            if any(token in row_label for token in ("margin", "yoy", "%")):
                continue
            for cell in row[1:]:
                value = _num(cell.value)
                assert value is None or not (0 < abs(value) < 0.001), (
                    f"GPRE {sheet_name} {cell.coordinate} still shows suspicious micro value {cell.value!r}"
                )
    wb.close()


def test_gpre_operating_drivers_excludes_spurious_future_quarters() -> None:
    wb = _load_model("GPRE")
    ws = wb["Operating_Drivers"]
    header = _find_row(ws, "Quarter")
    labels = [_text(ws.cell(header, cc).value) for cc in range(2, int(ws.max_column or 0) + 1)]
    assert "2027-Q1" not in labels
    assert "2029-Q2" not in labels
    assert "2049-Q3" not in labels
    assert "2026-Q1" in labels
    wb.close()


def test_gpre_promise_facility_qualification_remains_progress_not_completion() -> None:
    wb = _load_model("GPRE")
    ws = wb["Promise_Progress_UI"]
    row = _promise_row(ws, "2026-Q1 revisions", "45Z facility qualification")
    assert _text(ws.cell(row, 5).value) == ""
    progress = _text(ws.cell(row, 6).value)
    assert "All 8" in progress
    assert "operational" in progress.lower() or "qualifying" in progress.lower()
    assert _text(ws.cell(row, 7).value) == "On track"
    assert _text(ws.cell(row, 8).value) == "2026-Q1"
    assert _text(ws.cell(row, 9).value) == "2026-Q1"
    assert _text(ws.cell(row, 10).value) == "2026-03-31"
    assert "2049" not in " ".join(_text(ws.cell(row, cc).value) for cc in range(1, 12))
    wb.close()


def test_gpre_promise_source_backed_45z_and_cost_savings_values() -> None:
    wb = _load_model("GPRE")
    ws = wb["Promise_Progress_UI"]

    q4_45z = _promise_row(ws, "2025-Q4 revisions", "45Z monetization")
    assert _text(ws.cell(q4_45z, 4).value) == "Updated"
    assert _text(ws.cell(q4_45z, 5).value) == "$23.4m"
    assert _text(ws.cell(q4_45z, 6).value) == "YTD: $49.9m"
    assert _text(ws.cell(q4_45z, 7).value) == "Hit"
    assert _text(ws.cell(q4_45z, 8).value) == "2025-Q4"
    assert _text(ws.cell(q4_45z, 9).value) == "2025-Q4"
    assert _text(ws.cell(q4_45z, 10).value) == "2025-12-31"
    assert "YTD adds Q3 $26.5m and Q4 $23.4m" in _text(ws.cell(q4_45z, 11).value)

    expected_cost_rows = {
        "2024-Q4 revisions": ("$30m", "Executed: $30m"),
        "2025-Q1 revisions": ("$45m", "Remaining: $5m"),
        "2025-Q2 revisions": (">= $50m", "On pace to exceed $50m"),
    }
    for section, (actual, progress) in expected_cost_rows.items():
        row = _promise_row(ws, section, "Cost savings target")
        assert _text(ws.cell(row, 5).value) == actual
        assert _text(ws.cell(row, 6).value) == progress
        assert _text(ws.cell(row, 7).value) == "On track"
        assert _text(ws.cell(row, 8).value) == "Annualized program"

    wb.close()


def test_gpre_bs_segments_quarterly_total_assets_has_q2_2024_source_value() -> None:
    wb = _load_model("GPRE")
    ws = wb["BS_Segments"]
    header = _find_row(ws, "Quarter")
    col = _quarter_col(ws, header, "2024-Q2")
    row = _find_row(ws, "Total assets", start=header)
    assert _num(ws.cell(row, col).value) == pytest.approx(1763.6, abs=0.01)
    wb.close()


def test_gpre_bs_segments_current_maturities_use_debt_current_source_values() -> None:
    wb = _load_model("GPRE")
    ws = wb["BS_Segments"]
    header = _find_row(ws, "Quarter")
    row = _find_row(ws, "Current maturities of long-term debt", start=header)
    expected_values = {
        "2024-Q2": 1.830,
        "2024-Q3": 1.875,
        "2024-Q4": 2.118,
        "2025-Q1": 2.118,
        "2025-Q2": 2.125,
        "2025-Q3": 2.042,
        "2025-Q4": 3.924,
        "2026-Q1": 69.316,
    }
    for quarter, expected in expected_values.items():
        col = _quarter_col(ws, header, quarter)
        assert _num(ws.cell(row, col).value) == pytest.approx(expected, abs=0.001), (
            f"GPRE BS_Segments current maturities {quarter} should use debt_current XBRL source values"
        )
    wb.close()


def test_gpre_bs_segments_annual_total_assets_includes_2023_source_values() -> None:
    wb = _load_model("GPRE")
    ws = wb["BS_Segments"]
    annual_header = _find_row(ws, "Annual segments")
    year_row = _find_row(ws, "Year", start=annual_header)
    year_cols = {
        int(ws.cell(year_row, cc).value): cc
        for cc in range(2, ws.max_column + 1)
        if str(ws.cell(year_row, cc).value or "").isdigit()
    }
    assert {2023, 2024, 2025}.issubset(set(year_cols))

    total_assets_header = _find_row(ws, "Total assets", start=year_row)
    d_and_a_header = _find_row(ws, "Depreciation & amortization", start=year_row)
    operating_income_header = _find_row(ws, "Operating income (loss)", start=year_row)
    expected_2023 = {
        "Ethanol production": 1275.562,
        "Agribusiness and energy services": 413.937,
        "Corporate assets": 254.300,
        "Intersegment eliminations": -4.477,
    }
    component_sum = 0.0
    for segment, expected_value in expected_2023.items():
        row = _find_row(ws, segment, start=total_assets_header)
        value = _num(ws.cell(row, year_cols[2023]).value)
        assert value == pytest.approx(expected_value, abs=0.001), (
            f"GPRE annual Total assets 2023 {segment} should come from FY2024 10-K comparatives"
        )
        component_sum += float(value)
        assert _num(ws.cell(row, year_cols[2024]).value) is not None
        assert _num(ws.cell(row, year_cols[2025]).value) is not None
    assert component_sum == pytest.approx(1939.322, abs=0.001)

    d_and_a_labels = {
        _text(ws.cell(rr, 1).value)
        for rr in range(d_and_a_header + 1, operating_income_header)
        if _text(ws.cell(rr, 1).value)
    }
    operating_income_labels = {
        _text(ws.cell(rr, 1).value)
        for rr in range(operating_income_header + 1, total_assets_header)
        if _text(ws.cell(rr, 1).value)
    }
    total_asset_labels = {
        _text(ws.cell(rr, 1).value)
        for rr in range(total_assets_header + 1, min(total_assets_header + 12, int(ws.max_row or 0)) + 1)
        if _text(ws.cell(rr, 1).value)
    }
    assert "Corporate activities" in d_and_a_labels
    assert "Corporate activities" in operating_income_labels
    assert "Corporate activities" not in total_asset_labels
    assert "Corporate assets" in total_asset_labels
    wb.close()


def test_carbon_equipment_liabilities_render_rule_is_sector_or_value_specific() -> None:
    assert _should_render_carbon_equipment_liabilities("GPRE", {}) is True
    assert _should_render_carbon_equipment_liabilities("PBI", {}) is False
    assert _should_render_carbon_equipment_liabilities("ANF", {}) is False
    assert _should_render_carbon_equipment_liabilities("XYZ", {"2026-Q1": 12.3}) is True


def test_carbon_equipment_liabilities_visible_only_when_relevant() -> None:
    expected = {"PBI": False, "ANF": False, "GPRE": True}
    for ticker, should_render in expected.items():
        wb = _load_model(ticker)
        ws = wb["BS_Segments"]
        labels = {_text(ws.cell(rr, 1).value) for rr in range(1, int(ws.max_row or 0) + 1)}
        assert ("Carbon equipment liabilities" in labels) is should_render, (
            f"{ticker} should {'render' if should_render else 'suppress'} blank/irrelevant Carbon equipment liabilities"
        )
        wb.close()


def test_gpre_investment_case_45z_baseline_uses_reported_baseline_not_unknown() -> None:
    wb = _load_model("GPRE")
    wb_formula = load_workbook(_model_path("GPRE"), data_only=False, read_only=True, keep_vba=True)
    ws = wb["GPRE_Investment_Case"]
    ws_formula = wb_formula["GPRE_Investment_Case"]
    row = _find_row(ws, "Incremental 45Z uplift vs baseline")
    source_baseline = _num(ws.cell(28, 3).value)
    active_default = _num(ws.cell(28, 7).value) or _num(ws.cell(28, 4).value) or _num(ws.cell(28, 3).value)
    assert source_baseline is not None and source_baseline > 0
    assert active_default == pytest.approx(212.5, abs=0.1)
    assert "$C$28" in _text(ws_formula.cell(row, 2).value)
    assert "$G$28" in _text(ws_formula.cell(row, 3).value)
    assert f"C{row}-B{row}" in _text(ws_formula.cell(row, 4).value)
    wb_formula.close()
    wb.close()


def test_gpre_revolver_facility_size_is_not_below_reported_availability() -> None:
    wb = _load_model("GPRE")
    ws = wb["Valuation"]
    header = _find_row(ws, "Quarter")
    facility_row = _find_row(ws, "Revolver facility size", start=header)
    availability_row = _find_row(ws, "Revolver availability", start=header)
    for quarter in ("2023-Q4", "2025-Q4", "2026-Q1"):
        col = _quarter_col(ws, header, quarter)
        facility = _num(ws.cell(facility_row, col).value)
        availability = _num(ws.cell(availability_row, col).value)
        assert facility is None or availability is None or facility + 0.001 >= availability, (
            f"GPRE revolver facility size {quarter} should not be below availability"
        )
    wb.close()


def test_anf_bs_driver_yoy_uses_hidden_prior_quarters_for_visible_window() -> None:
    wb = _load_model_with_styles("ANF")
    ws = wb["BS_Segments"]
    header = _find_row(ws, "Quarter")
    windows = _anf_visible_windows(wb)
    visible_quarters = windows["bs_segments"]
    checked_quarters = visible_quarters[:4]
    _assert_prior_year_history_and_facts_exist(wb, checked_quarters, visible_quarters)
    for row_label in ("Inventory YoY", "Sales YoY", "Diluted shares YoY"):
        row = _find_row(ws, row_label, start=header)
        for quarter in checked_quarters:
            col = _quarter_col(ws, header, quarter)
            assert _num(ws.cell(row, col).value) is not None, (
                f"ANF BS_Segments {row_label} {quarter} should use hidden prior-year quarters"
            )
            assert _has_bucket_fill(ws.cell(row, col)), (
                f"ANF BS_Segments {row_label} {quarter} should retain source-backed comparison fill"
            )
    wb.close()


def test_anf_older_promise_progressions_have_source_backed_actuals() -> None:
    wb = _load_model("ANF")
    ws = wb["Promise_Progress_UI"]
    checks = [
        ("2024 guidance progression", "Capex", "Mixed"),
        ("2023 guidance progression", "Q1 sales growth", "Hit"),
        ("2023 guidance progression", "Capex", "Hit"),
        ("2022 guidance progression", "Net sales growth", "Mixed"),
    ]
    for section, metric, status in checks:
        row = _promise_row(ws, section, metric)
        assert _text(ws.cell(row, 7).value), f"{section} {metric} should have Actual"
        assert _text(ws.cell(row, 8).value) == status
    wb.close()
