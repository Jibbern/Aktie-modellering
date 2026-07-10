from __future__ import annotations

import hashlib
import json
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def _data_root() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = parent / "StockModelData"
        if candidate.exists():
            return candidate
    return ROOT.parent / "StockModelData"


DATA_ROOT = _data_root()
ARCHITECTURE_PATH = ROOT / "docs" / "workbook_block_architecture.json"
BINDING_MAP_PATH = ROOT / "docs" / "workbook_binding_map.json"
LAB_PATH = ROOT / "templates" / "lab" / "ANF_template_lab.xlsx"

STANDARD_VISIBLE_SHEETS = {
    "SUMMARY",
    "Valuation",
    "BS_Segments",
    "Operating_Drivers",
    "{ticker}_Investment_Case",
    "Quarter_Notes_UI",
    "Promise_Progress_UI",
    "QA_Log",
    "Needs_Review",
    "QA_Checks",
}

REQUIRED_BLOCK_KEYS = {
    "block_id",
    "sheet",
    "range",
    "title_header_cells",
    "static_label_cells",
    "writable_value_cells",
    "formula_cells",
    "hidden_helper_cells",
    "normalized_fields",
    "support_sheets_used",
    "current_code_owner",
    "future_intended_owner",
    "required_fields",
    "optional_fields",
    "source_policy",
    "missing_data_behavior",
    "validation_rules",
    "standardization_status",
}

COMPANY_SPECIFIC_TERMS = (
    "Abercrombie",
    "Hollister",
    "Pitney Bowes",
    "Presort",
    "SendTech",
    "Green Plains",
    "45Z",
    "RIN",
    "crush margin",
)

SUMMARY_VALUE_CELLS_THAT_MUST_NOT_BE_STATIC = {"B9", "B10", "B11", "B30", "B32", "B36"}


def _architecture() -> dict:
    return json.loads(ARCHITECTURE_PATH.read_text(encoding="utf-8"))


def _binding_map() -> dict:
    return json.loads(BINDING_MAP_PATH.read_text(encoding="utf-8"))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _col_to_int(col: str) -> int:
    value = 0
    for char in col:
        value = value * 26 + (ord(char.upper()) - ord("A") + 1)
    return value


def _coord_to_tuple(coord: str) -> tuple[int, int]:
    col = "".join(ch for ch in coord if ch.isalpha())
    row = "".join(ch for ch in coord if ch.isdigit())
    return _col_to_int(col), int(row)


def _parse_range(target: str) -> tuple[int, int, int, int]:
    start, end = target.split(":", 1)
    left, top = _coord_to_tuple(start)
    right, bottom = _coord_to_tuple(end)
    return left, top, right, bottom


def _cell_in_range(coord: str, target: str) -> bool:
    col, row = _coord_to_tuple(coord)
    left, top, right, bottom = _parse_range(target)
    return left <= col <= right and top <= row <= bottom


def _looks_numeric(value: object) -> bool:
    if isinstance(value, (int, float)):
        return True
    text = str(value or "").strip().replace(",", ".").replace("%", "")
    if not text or text.upper() == "N/A":
        return False
    try:
        float(text)
    except ValueError:
        return False
    return True


def test_template_lab_is_byte_identical_macro_free_copy_of_anf_source() -> None:
    assert LAB_PATH.exists()
    assert LAB_PATH.suffix.lower() == ".xlsx"

    with zipfile.ZipFile(LAB_PATH) as zf:
        assert not any(name.lower().endswith("vbaproject.bin") for name in zf.namelist())

    payload = _architecture()
    lab_meta = payload["template_lab"]
    source_path = Path(lab_meta["source_path"])

    assert source_path == DATA_ROOT / "outputs" / "Excel stock models" / "ANF_model.xlsx"
    assert lab_meta["lab_path"] == str(LAB_PATH)
    assert lab_meta["byte_identical"] is True
    assert lab_meta["source_sha256"] == _sha256(source_path)
    assert lab_meta["lab_sha256"] == _sha256(LAB_PATH)
    assert lab_meta["source_sha256"] == lab_meta["lab_sha256"]


def test_template_lab_loads_with_standard_visible_sheet_family() -> None:
    wb = load_workbook(LAB_PATH, read_only=True, data_only=False)
    try:
        sheetnames = set(wb.sheetnames)
        expected = {sheet.replace("{ticker}", "ANF") for sheet in STANDARD_VISIBLE_SHEETS}

        assert expected <= sheetnames
    finally:
        wb.close()


def test_workbook_block_architecture_covers_standard_visible_sheets() -> None:
    payload = _architecture()

    assert {"version", "generated_at", "source_workbooks", "template_lab", "standard_visible_sheets", "blocks"} <= set(payload)
    assert set(payload["standard_visible_sheets"]) == STANDARD_VISIBLE_SHEETS

    blocks = payload["blocks"]
    assert blocks
    assert {block["sheet"] for block in blocks} >= STANDARD_VISIBLE_SHEETS

    for block in blocks:
        assert REQUIRED_BLOCK_KEYS <= set(block)
        assert block["block_id"]
        assert block["sheet"] in STANDARD_VISIBLE_SHEETS
        assert block["range"]
        assert block["standardization_status"] in {"standard", "sector_specific", "ticker_specific"}
        assert block["source_policy"] in {"source-backed", "profile-backed", "manual", "derived", "validation-output", "mixed"}
        assert block["missing_data_behavior"]
        assert isinstance(block["title_header_cells"], list)
        assert isinstance(block["static_label_cells"], list)
        assert isinstance(block["writable_value_cells"], list)
        assert isinstance(block["formula_cells"], list)
        assert isinstance(block["hidden_helper_cells"], list)


def test_required_blocks_have_normalized_fields_and_missing_data_behavior() -> None:
    required_blocks = [
        block
        for block in _architecture()["blocks"]
        if block["required_fields"] or any(binding.get("required") for binding in block.get("bindings", []))
    ]

    assert required_blocks
    for block in required_blocks:
        assert block["normalized_fields"]
        assert block["missing_data_behavior"]
        assert block["validation_rules"]


def test_no_company_specific_terms_are_standard_template_labels() -> None:
    offenders: list[str] = []
    for block in _architecture()["blocks"]:
        if block["standardization_status"] != "standard":
            continue
        label_cells = [*block["title_header_cells"], *block["static_label_cells"]]
        for label in label_cells:
            value = str(label.get("value", ""))
            for term in COMPANY_SPECIFIC_TERMS:
                if term.lower() in value.lower():
                    offenders.append(f"{block['block_id']} {label.get('cell')}: {value}")

    assert offenders == []


def test_static_label_cells_never_overlap_writable_binding_targets() -> None:
    offenders: list[str] = []
    for block in _architecture()["blocks"]:
        binding_targets = [binding["target"] for binding in block.get("bindings", [])]
        for label in block["static_label_cells"]:
            cell = label["cell"]
            for target in binding_targets:
                if _cell_in_range(cell, target):
                    offenders.append(f"{block['block_id']} {cell} inside {target}: {label.get('value')}")

    assert offenders == []


def test_static_label_cells_do_not_hold_writable_numeric_examples() -> None:
    offenders: list[str] = []
    for block in _architecture()["blocks"]:
        binding_targets = [binding["target"] for binding in block.get("bindings", [])]
        for label in block["static_label_cells"]:
            if not _looks_numeric(label.get("value")):
                continue
            cell = label["cell"]
            if any(_cell_in_range(cell, target) for target in binding_targets):
                offenders.append(f"{block['block_id']} {cell}: {label.get('value')}")

    assert offenders == []


def test_summary_known_value_cells_are_not_static_labels() -> None:
    offenders = []
    for block in _architecture()["blocks"]:
        if block["sheet"] != "SUMMARY":
            continue
        for label in block["static_label_cells"]:
            if label["cell"] in SUMMARY_VALUE_CELLS_THAT_MUST_NOT_BE_STATIC:
                offenders.append(f"{block['block_id']} {label['cell']}: {label.get('value')}")

    assert offenders == []


def test_blocks_classify_writable_and_company_specific_examples_separately() -> None:
    for block in _architecture()["blocks"]:
        assert "template_label_cells" in block
        assert "row_label_cells" in block
        assert "writable_example_cells" in block
        assert "company_specific_example_cells" in block
        assert "source_specific_example_cells" in block
        for binding in block.get("bindings", []):
            target = binding["target"]
            assert any(example["target"] == target for example in block["writable_example_cells"])


def test_every_writable_binding_resolves_to_a_block() -> None:
    bindings = [entry for entry in _binding_map()["bindings"] if entry["writable"]]
    writable_binding_ids = {entry["binding_id"] for entry in bindings}
    architecture_blocks = {block["block_id"] for block in _architecture()["blocks"]}
    block_binding_ids = {
        binding["binding_id"]
        for block in _architecture()["blocks"]
        for binding in block.get("bindings", [])
    }
    direct_block_ids = {
        entry["binding_id"]
        for entry in bindings
        if entry.get("block_id") in architecture_blocks
    }

    assert writable_binding_ids <= block_binding_ids | direct_block_ids


def test_guardrails_do_not_create_gtx_or_macro_outputs() -> None:
    forbidden = [
        DATA_ROOT / "outputs" / "stress_tests" / "GTX_new_ticker_engine" / "GTX_model.xlsx",
        DATA_ROOT / "outputs" / "Excel stock models" / "GTX_model.xlsx",
        DATA_ROOT / "outputs" / "Excel stock models" / "GTX_model.xlsm",
    ]

    assert [str(path) for path in forbidden if path.exists()] == []
    assert list((ROOT / "templates").rglob("*.xlsm")) == []
