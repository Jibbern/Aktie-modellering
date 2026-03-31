from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from pbi_xbrl.debt_parser import parse_debt_tranches_from_primary_doc
from pbi_xbrl.workbook_gap_audit import (
    STATUS_NOT_FOUND,
    STATUS_PARSE_FAILED,
    STATUS_WRITE_FAILED,
    build_workbook_gap_rows,
    classify_gap_status,
    infer_expected_source_family,
)


PIPELINE_EMPTY_DF = pd.DataFrame()


def _empty_pipeline_bundle(hist: pd.DataFrame, debt_tranches: pd.DataFrame, debt_tranches_latest: pd.DataFrame) -> tuple:
    return (
        hist,
        PIPELINE_EMPTY_DF,
        debt_tranches,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        debt_tranches_latest,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        PIPELINE_EMPTY_DF,
        {},
    )


def _write_minimal_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation"
    ws.cell(246, 12, "Convertible notes")
    ws.cell(248, 12, "No convertible debt identified in latest debt set.")
    for sheet_name in [
        "Quarter_Notes_UI",
        "Needs_Review",
        "QA_Log",
        "Operating_Drivers",
        "Promise_Progress_UI",
        "BS_Segments",
    ]:
        wb.create_sheet(sheet_name)
    qn = wb["Quarter_Notes_UI"]
    qn["A1"] = "Generated at test"
    nr = wb["Needs_Review"]
    nr.append(["priority", "issue_family", "severity", "first_seen_q", "last_seen_q", "quarter_count", "latest_message", "recommended_action", "source"])
    qa = wb["QA_Log"]
    qa["A1"] = "quarter"
    od = wb["Operating_Drivers"]
    od["A1"] = "Metric / segment"
    wb.save(path)
    wb.close()


def _write_convertible_visible_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation"
    ws.cell(246, 12, "Convertible notes")
    ws.cell(248, 12, "2.25% notes due 2027")
    for sheet_name in [
        "Quarter_Notes_UI",
        "Needs_Review",
        "QA_Log",
        "Operating_Drivers",
        "Promise_Progress_UI",
        "BS_Segments",
        "Economics_Overlay",
    ]:
        wb.create_sheet(sheet_name)
    wb.save(path)
    wb.close()


def test_classify_gap_status_values() -> None:
    assert classify_gap_status(source_exists=False, parse_ok=False, write_ok=False) == STATUS_NOT_FOUND
    assert classify_gap_status(source_exists=True, parse_ok=False, write_ok=False) == STATUS_PARSE_FAILED
    assert classify_gap_status(source_exists=True, parse_ok=True, write_ok=False) == STATUS_WRITE_FAILED


def test_infer_expected_source_family_maps_debt_segment_and_notes() -> None:
    assert infer_expected_source_family(sheet="Valuation", row_label="Convertible notes").startswith("debt_tranches")
    assert infer_expected_source_family(sheet="Operating_Drivers", row_label="segment_support").startswith("historical_segment")
    assert infer_expected_source_family(sheet="Quarter_Notes_UI", row_label="note").startswith("earnings_release")


def test_parse_debt_tranches_rescues_adjacent_numeric_amount_from_rate_like_cell() -> None:
    html = b"""
    <html><body>
      <p>(dollars in thousands)</p>
      <table>
        <tr><th>Debt obligations</th><th>2025</th><th>2026</th></tr>
        <tr><td>Convertible Notes due August 2030</td><td>1.50</td><td>230,000</td></tr>
        <tr><td>Notes due March 2027</td><td>6.875</td><td>346,700</td></tr>
        <tr><td>Total debt</td><td>1.99</td><td>1,993,038</td></tr>
      </table>
    </body></html>
    """
    rows, score, table_total_debt, *_ = parse_debt_tranches_from_primary_doc(html, quarter_end=pd.Timestamp("2025-12-31").date())
    assert score > 0
    assert table_total_debt is not None
    conv = next(row for row in rows if "Convertible" in str(row.get("name") or ""))
    assert conv["amount"] == 230000000.0
    assert conv["amount_col_idx"] == 2
    assert conv["parse_quality"] in {"asof_matched", "asof_matched_adjacent_numeric"}


def test_build_workbook_gap_rows_flags_blank_convertible_block_when_raw_source_exists(tmp_path: Path) -> None:
    repo_root = tmp_path
    (repo_root / "Excel stock models").mkdir(parents=True, exist_ok=True)
    (repo_root / "sec_cache" / "PBI" / "pipeline_bundle_cache").mkdir(parents=True, exist_ok=True)
    workbook_path = repo_root / "Excel stock models" / "PBI_model.xlsx"
    _write_minimal_workbook(workbook_path)

    hist = pd.DataFrame([{"quarter": pd.Timestamp("2025-12-31")}])
    debt_tranches = pd.DataFrame(
        [
            {
                "quarter": pd.Timestamp("2025-12-31"),
                "tranche_name": "Convertible Notes due August 2030",
                "amount": 230000000.0,
                "doc": "pbi-20251231.htm",
                "form": "10-K",
                "accn": "0001628280-26-009650",
            }
        ]
    )
    debt_tranches_latest = pd.DataFrame()
    pd.to_pickle(
        _empty_pipeline_bundle(hist, debt_tranches, debt_tranches_latest),
        repo_root / "sec_cache" / "PBI" / "pipeline_bundle_cache" / "PBI.pkl",
    )

    rows = build_workbook_gap_rows(repo_root, "PBI")
    conv_rows = [row for row in rows if row.sheet == "Valuation" and row.row_label == "Convertible notes"]
    assert conv_rows
    assert conv_rows[0].status == STATUS_PARSE_FAILED
    assert "latest debt view is empty" in conv_rows[0].reason


def test_build_workbook_gap_rows_flags_visible_convertible_row_when_conversion_terms_missing(tmp_path: Path) -> None:
    repo_root = tmp_path
    (repo_root / "Excel stock models").mkdir(parents=True, exist_ok=True)
    (repo_root / "sec_cache" / "GPRE" / "pipeline_bundle_cache").mkdir(parents=True, exist_ok=True)
    workbook_path = repo_root / "Excel stock models" / "GPRE_model.xlsx"
    _write_convertible_visible_workbook(workbook_path)

    hist = pd.DataFrame([{"quarter": pd.Timestamp("2025-12-31")}])
    debt_tranches = pd.DataFrame()
    debt_tranches_latest = pd.DataFrame(
        [
            {
                "quarter": pd.Timestamp("2025-12-31"),
                "tranche_name": "2.25 % convertible notes due 2027",
                "instrument_type": "convertible",
                "coupon_pct": 2.25,
                "maturity_year": 2027,
                "doc": "gpre-20251231.htm",
            }
        ]
    )
    pd.to_pickle(
        _empty_pipeline_bundle(hist, debt_tranches, debt_tranches_latest),
        repo_root / "sec_cache" / "GPRE" / "pipeline_bundle_cache" / "GPRE.pkl",
    )

    rows = build_workbook_gap_rows(repo_root, "GPRE")
    conv_rows = [row for row in rows if row.sheet == "Valuation" and "2027" in row.row_label]
    assert conv_rows
    assert conv_rows[0].status == STATUS_PARSE_FAILED
    assert "parsed conversion terms are still missing" in conv_rows[0].reason


def test_build_workbook_gap_rows_flags_write_failure_when_conversion_terms_exist_but_cells_blank(tmp_path: Path) -> None:
    repo_root = tmp_path
    (repo_root / "Excel stock models").mkdir(parents=True, exist_ok=True)
    (repo_root / "sec_cache" / "GPRE" / "pipeline_bundle_cache").mkdir(parents=True, exist_ok=True)
    workbook_path = repo_root / "Excel stock models" / "GPRE_model.xlsx"
    _write_convertible_visible_workbook(workbook_path)

    hist = pd.DataFrame([{"quarter": pd.Timestamp("2025-12-31")}])
    debt_tranches = pd.DataFrame()
    debt_tranches_latest = pd.DataFrame(
        [
            {
                "quarter": pd.Timestamp("2025-12-31"),
                "tranche_name": "2.25 % convertible notes due 2027",
                "instrument_type": "convertible",
                "coupon_pct": 2.25,
                "maturity_year": 2027,
                "conversion_price": 31.62,
                "shares_on_full_conversion": 1.897,
                "conversion_terms_source": "gpre-20251231.htm",
            }
        ]
    )
    pd.to_pickle(
        _empty_pipeline_bundle(hist, debt_tranches, debt_tranches_latest),
        repo_root / "sec_cache" / "GPRE" / "pipeline_bundle_cache" / "GPRE.pkl",
    )

    rows = build_workbook_gap_rows(repo_root, "GPRE")
    conv_rows = [row for row in rows if row.sheet == "Valuation" and "2027" in row.row_label]
    assert conv_rows
    assert conv_rows[0].status == STATUS_WRITE_FAILED
    assert "missing conversion price, shares on full conversion" in conv_rows[0].reason
