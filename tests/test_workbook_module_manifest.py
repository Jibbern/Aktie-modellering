from __future__ import annotations

from collections import Counter
from copy import deepcopy
import json
from pathlib import Path

from openpyxl import load_workbook
import pytest

from pbi_xbrl.json_schema_validation import load_json_strict, validate_json_schema
from pbi_xbrl.standard_template_formula_contract import formula_target_contracts
from pbi_xbrl.standard_template_shell_identity import verify_shell_identity
from pbi_xbrl.workbook_modules import (
    apply_runtime_sheet_order,
    binding_owners,
    build_profile_binding_payload,
    build_profile_shell_manifest,
    canonical_json_sha256,
    load_workbook_module_manifest,
    profile_id_for_ticker,
    resolve_module_profile,
    resolve_ticker_module_route,
    sheet_contracts,
    validate_binding_module_ownership,
    validate_workbook_module_manifest,
    validate_workbook_execution_ownership,
)
from scripts.materialize_standard_template_shell import materialize_shell
from scripts.materialize_standard_template_shell import _configure_investment_case_ownership_zones


ROOT = Path(__file__).resolve().parents[1]
MODULE_MANIFEST = ROOT / "docs" / "workbook_module_manifest.json"
MODULE_SCHEMA = ROOT / "docs" / "workbook_module_manifest.schema.json"
SHELL_MANIFEST = ROOT / "docs" / "standard_template_shell_manifest.json"
BINDING_MAP = ROOT / "docs" / "workbook_binding_map.json"


def _payload() -> dict:
    return load_workbook_module_manifest(MODULE_MANIFEST)


def _legacy_anf_path() -> Path:
    for parent in (ROOT, *ROOT.parents):
        candidate = parent / "StockModelData" / "outputs" / "Excel stock models" / "ANF_model.xlsx"
        if candidate.exists():
            return candidate
    raise AssertionError("Read-only ANF legacy oracle is unavailable.")


def test_module_manifest_schema_and_semantics_pass() -> None:
    payload = load_json_strict(MODULE_MANIFEST)

    assert validate_json_schema(payload, load_json_strict(MODULE_SCHEMA)) == []
    assert validate_workbook_module_manifest(payload) == []
    assert len(payload["modules"]) == 12
    assert len(payload["union_sheet_order"]) == 44


def test_debt_product_layout_is_bounded_in_the_module_manifest() -> None:
    payload = _payload()
    debt = next(module for module in payload["modules"] if module["module_id"] == "debt_liquidity")
    sheets = {row["sheet"]: row for row in debt["sheets"]}

    expected = {
        "Debt_Profile": (32, 32, ["B", "C", "H", "I", "J"], 95),
        "Revolver_History": (36, 32, ["N", "O", "P"], 90),
        "Leverage_Liquidity": (36, 48, ["L", "M", "N"], 90),
        "Debt_Credit_Notes": (34, 48, ["A", "E", "G", "H"], 95),
    }
    for sheet_name, (header, body, wrap, zoom) in expected.items():
        assert sheets[sheet_name]["header_row_height"] == header
        assert sheets[sheet_name]["body_row_height"] == body
        assert sheets[sheet_name]["wrap_columns"] == wrap
        assert sheets[sheet_name]["zoom_scale"] == zoom


def test_legacy_inventory_explicitly_classifies_retired_duplicate_valuation_sheets() -> None:
    payload = _payload()
    rows = payload["legacy_sheet_inventory"]
    wb = load_workbook(_legacy_anf_path(), read_only=True, data_only=False)
    try:
        physical_sheets = set(wb.sheetnames)
    finally:
        wb.close()

    inventory_sheets = {row["legacy_sheet"] for row in rows}
    retired_sheets = {"Valuation_Summary", "Valuation_Grid"}
    assert inventory_sheets == physical_sheets

    retired_rows = {
        row["legacy_sheet"]: row
        for row in rows
        if row["legacy_sheet"] in retired_sheets
    }
    assert set(retired_rows) == retired_sheets
    assert all(row["disposition"] == "rejected_redundant" for row in retired_rows.values())
    assert all(row["replacement"] == "Valuation" for row in retired_rows.values())

    assert Counter(row["legacy_class"] for row in rows) == {
        "A": 10,
        "B": 12,
        "C": 15,
        "D": 10,
        "E": 7,
        "F": 3,
    }
    slides_guidance = next(row for row in rows if row["legacy_sheet"] == "Slides_Guidance")
    assert slides_guidance["disposition"] == "rejected_redundant"
    assert slides_guidance["replacement"] == "Guidance_Normalized"
    assert all(row.get("reason") for row in rows)


def test_every_executable_binding_has_exactly_one_module_owner() -> None:
    payload = _payload()
    bindings = load_json_strict(BINDING_MAP)
    owners = binding_owners(payload)

    assert validate_binding_module_ownership(payload, bindings) == []
    assert len(owners) == len(bindings["bindings"])
    assert {row["binding_id"] for row in bindings["bindings"]} == set(owners)


def test_profiles_are_explicit_dependency_closed_and_pack_declarative() -> None:
    payload = _payload()

    assert profile_id_for_ticker(payload, "ANF") == "anf"
    assert profile_id_for_ticker(payload, "PBI") == "pbi"
    assert profile_id_for_ticker(payload, "GPRE") == "gpre"
    assert resolve_module_profile(payload, "anf").profile_pack_ids == ("retail_operating_pack",)
    assert resolve_module_profile(payload, "pbi").profile_pack_ids == ("shipping_mail_pack", "bank_pack")
    assert resolve_module_profile(payload, "gpre").profile_pack_ids == ("commodity_ethanol_pack",)
    bindings = load_json_strict(BINDING_MAP)
    anf_binding = build_profile_binding_payload(bindings, payload, resolve_module_profile(payload, "anf"))
    anf_scenario_packs = {row["profile_pack_id"]: row["scenario_driver_ids"] for row in anf_binding["scenario_profile_packs"]}
    assert set(anf_scenario_packs) == {"retail_operating_pack"}
    assert {
        "revenue_growth",
        "comparable_sales",
        "store_openings",
        "store_closures",
        "store_remodels",
        "aur_pricing",
        "inventory_change",
        "tariff_impact",
        "operating_margin_guidance",
        "capital_expenditures_guidance",
        "adjusted_eps_guidance",
    } == set(anf_scenario_packs["retail_operating_pack"])

    profiles = {row["profile_id"]: row for row in payload["profiles"]}
    modules = {row["module_id"]: row for row in payload["modules"]}
    for profile in profiles.values():
        enabled = set(profile["enabled_modules"])
        for module_id in enabled:
            assert set(modules[module_id]["dependencies"]) <= enabled

    broken = deepcopy(payload)
    core_only = next(row for row in broken["profiles"] if row["profile_id"] == "core_only")
    core_only["enabled_modules"].remove("core_financial_history")
    assert any("without dependencies" in issue for issue in validate_workbook_module_manifest(broken))


def test_commodity_pack_owns_derivative_runtime_sheets_without_cross_ticker_inheritance() -> None:
    payload = _payload()
    gpre = resolve_ticker_module_route(payload, "GPRE").resolved_profile
    assert gpre is not None
    assert set(gpre.owned_runtime_sheets) == {
        "Derivative_OCI_Bridge",
        "Derivative_Crush_Tests",
    }
    assert set(gpre.visible_runtime_sheets) == set(gpre.owned_runtime_sheets)
    assert gpre.runtime_sheet_states == {
        "Derivative_Crush_Tests": "visible",
        "Derivative_OCI_Bridge": "visible",
    }

    ordered = apply_runtime_sheet_order(
        ("Promise_Progress_UI", "Basis_Proxy_Sandbox", "Hidden_Value_Flags"),
        gpre.ordered_runtime_sheets,
    )
    assert ordered == (
        "Promise_Progress_UI",
        "Derivative_OCI_Bridge",
        "Basis_Proxy_Sandbox",
        "Derivative_Crush_Tests",
        "Hidden_Value_Flags",
    )

    for ticker in ("PBI", "ANF"):
        resolved = resolve_ticker_module_route(payload, ticker).resolved_profile
        assert resolved is not None
        assert resolved.owned_runtime_sheets == ()
        assert resolved.visible_runtime_sheets == ()
        assert resolved.ordered_runtime_sheets == ()
    assert resolve_ticker_module_route(payload, "FRESHCO").resolved_profile is None


def test_runtime_sheet_ownership_is_independent_of_visibility_order_and_source_order() -> None:
    payload = _payload()
    baseline = resolve_module_profile(payload, "gpre")

    hidden = deepcopy(payload)
    commodity = next(
        row for row in hidden["profile_packs"] if row["pack_id"] == "commodity_ethanol_pack"
    )
    commodity["runtime_sheets"] = list(reversed(commodity["runtime_sheets"]))
    for row in commodity["runtime_sheets"]:
        row["visibility"] = "hidden"
    resolved_hidden = resolve_module_profile(hidden, "gpre")
    assert resolved_hidden.owned_runtime_sheets == baseline.owned_runtime_sheets
    assert resolved_hidden.visible_runtime_sheets == ()
    assert resolved_hidden.runtime_sheet_states == {
        "Derivative_Crush_Tests": "hidden",
        "Derivative_OCI_Bridge": "hidden",
    }
    base_order = ("Promise_Progress_UI", "Basis_Proxy_Sandbox", "Hidden_Value_Flags")
    assert apply_runtime_sheet_order(
        base_order,
        resolved_hidden.ordered_runtime_sheets,
    ) == apply_runtime_sheet_order(
        base_order,
        baseline.ordered_runtime_sheets,
    )

    missing_anchor = tuple(
        {**row, "order_after": "Missing_Anchor"}
        if row["sheet"] == "Derivative_OCI_Bridge"
        else row
        for row in baseline.ordered_runtime_sheets
    )
    with pytest.raises(ValueError, match="unknown anchor or cycle"):
        apply_runtime_sheet_order(base_order, missing_anchor)


def test_runtime_sheet_duplicate_or_conflicting_pack_ownership_fails_closed() -> None:
    payload = _payload()
    commodity = next(
        row for row in payload["profile_packs"] if row["pack_id"] == "commodity_ethanol_pack"
    )
    duplicate = deepcopy(payload)
    duplicate_commodity = next(
        row for row in duplicate["profile_packs"] if row["pack_id"] == "commodity_ethanol_pack"
    )
    duplicate_commodity["runtime_sheets"].append(
        deepcopy(duplicate_commodity["runtime_sheets"][0])
    )
    assert any(
        "Duplicate runtime sheet" in issue
        for issue in validate_workbook_module_manifest(duplicate)
    )

    conflict = deepcopy(payload)
    bank_pack = next(row for row in conflict["profile_packs"] if row["pack_id"] == "bank_pack")
    bank_pack["runtime_sheets"].append(deepcopy(commodity["runtime_sheets"][0]))
    assert any(
        "owned by both profile packs" in issue
        for issue in validate_workbook_module_manifest(conflict)
    )


def test_ticker_module_route_is_explicit_deterministic_and_fail_closed() -> None:
    payload = _payload()

    assert resolve_ticker_module_route(payload, "GPRE").profile_pack_ids == (
        "commodity_ethanol_pack",
    )
    assert resolve_ticker_module_route(payload, "PBI").profile_pack_ids == (
        "shipping_mail_pack",
        "bank_pack",
    )
    assert resolve_ticker_module_route(payload, "ANF").profile_pack_ids == (
        "retail_operating_pack",
    )

    unsupported = resolve_ticker_module_route(payload, "FRESHCO")
    assert unsupported.status == "unsupported"
    assert unsupported.profile_id == ""
    assert unsupported.profile_pack_ids == ()
    assert unsupported.resolved_profile is None

    declared = resolve_ticker_module_route(
        payload,
        "FRESHCO",
        declared_profile_id="core_only",
    )
    assert declared.status == "resolved_declared_profile"
    assert declared.profile_id == "core_only"
    assert declared.profile_pack_ids == ()

    declared_gpre = resolve_ticker_module_route(
        payload,
        "FRESHCO",
        declared_profile_id="gpre",
    )
    assert declared_gpre.status == "resolved_declared_profile"
    assert declared_gpre.profile_pack_ids == ("commodity_ethanol_pack",)

    reordered = deepcopy(payload)
    reordered["profiles"] = list(reversed(reordered["profiles"]))
    reordered["ticker_profile_map"] = dict(reversed(list(reordered["ticker_profile_map"].items())))
    assert resolve_ticker_module_route(reordered, "GPRE").to_dict() == resolve_ticker_module_route(
        payload,
        "GPRE",
    ).to_dict()
    assert resolve_ticker_module_route(reordered, "FRESHCO").to_dict() == unsupported.to_dict()


def test_ticker_module_route_rejects_unknown_conflicting_and_ambiguous_profiles() -> None:
    payload = _payload()

    with pytest.raises(ValueError, match="Unknown workbook module profile"):
        resolve_ticker_module_route(payload, "FRESHCO", declared_profile_id="missing_profile")
    with pytest.raises(ValueError, match="conflicts with declared profile"):
        resolve_ticker_module_route(payload, "GPRE", declared_profile_id="pbi")

    duplicate_profile = deepcopy(payload)
    duplicate_profile["profiles"].append(deepcopy(duplicate_profile["profiles"][-1]))
    with pytest.raises(ValueError, match="Duplicate profile_id"):
        resolve_ticker_module_route(duplicate_profile, "GPRE")

    ambiguous_ticker = deepcopy(payload)
    ambiguous_ticker["ticker_profile_map"]["freshco"] = "core_only"
    ambiguous_ticker["ticker_profile_map"]["FRESHCO"] = "full_union"
    with pytest.raises(ValueError, match="Ambiguous workbook module ticker registration"):
        resolve_ticker_module_route(ambiguous_ticker, "FRESHCO")


def test_generic_shared_block_requires_sheet_owner_dependency() -> None:
    broken = deepcopy(_payload())
    non_gaap = next(row for row in broken["modules"] if row["module_id"] == "non_gaap_adjustments")
    non_gaap["dependencies"].remove("balance_cash_flow")
    non_gaap["visible_blocks"].append(
        {
            "block_id": "synthetic_shared_balance_sheet_block",
            "sheet": "BS_Segments",
            "target": "A79:I80",
        }
    )
    core_only = next(row for row in broken["profiles"] if row["profile_id"] == "core_only")
    core_only["enabled_modules"] = ["core_financial_history", "non_gaap_adjustments", "qa_lineage"]

    issues = validate_workbook_module_manifest(broken)

    assert any(
        "Module 'non_gaap_adjustments' owns visible block 'synthetic_shared_balance_sheet_block' on sheet "
        "'BS_Segments', owned by 'balance_cash_flow', without a direct or transitive dependency."
        in issue
        for issue in issues
    )


def test_formula_name_block_pack_and_dimension_ownership_fail_closed() -> None:
    payload = _payload()

    duplicate_formula = deepcopy(payload)
    duplicate_formula["modules"][1]["formula_ids"].append("revenue_ttm")
    assert any("Duplicate formula_id 'revenue_ttm'" in issue for issue in validate_workbook_module_manifest(duplicate_formula))

    unknown_formula = deepcopy(payload)
    unknown_formula["modules"][0]["formula_ids"].append("unknown_formula_contract")
    assert any("Formula ownership mismatch" in issue for issue in validate_workbook_module_manifest(unknown_formula))

    duplicate_name = deepcopy(payload)
    duplicate_name["modules"][1]["defined_name_ids"].append("summary_key_financials_anchor")
    assert any("Duplicate defined_name_id" in issue for issue in validate_workbook_module_manifest(duplicate_name))

    overlapping_blocks = deepcopy(payload)
    operating = next(row for row in overlapping_blocks["modules"] if row["module_id"] == "operating_drivers")
    operating["visible_blocks"][1]["target"] = "A50:N125"
    operating["style_ownership"][1]["target"] = "A50:N125"
    assert any("overlap on 'Operating_Drivers'" in issue for issue in validate_workbook_module_manifest(overlapping_blocks))

    invalid_host = deepcopy(payload)
    invalid_host["profile_packs"][0]["host_module_id"] = "core_financial_history"
    assert any("is not a profile_pack_host" in issue for issue in validate_workbook_module_manifest(invalid_host))

    duplicate_driver = deepcopy(payload)
    duplicate_driver["profile_packs"][0]["scenario_driver_ids"].append("revenue_growth")
    assert any("Duplicate scenario_driver_id" in issue for issue in validate_workbook_module_manifest(duplicate_driver))

    invalid_dimensions = deepcopy(payload)
    core_profile = next(row for row in invalid_dimensions["profiles"] if row["profile_id"] == "core_only")
    core_profile["dimensions"].extend(
        [
            {"dimension_id": "total_company", "display_name": "Duplicate", "members_source": "universal"},
            {"dimension_id": "invented_dimension", "display_name": "Invalid", "members_source": "universal"},
        ]
    )
    issues = validate_workbook_module_manifest(invalid_dimensions)
    assert any("Duplicate dimension_id" in issue for issue in issues)
    assert any("unknown dimensions ['invented_dimension']" in issue for issue in issues)


def test_every_actual_formula_contract_has_one_owner_and_owned_target() -> None:
    payload = _payload()
    formula_ids = [formula_id for module in payload["modules"] for formula_id in module["formula_ids"]]

    assert Counter(formula_ids) == Counter(contract.formula_id for contract in formula_target_contracts())
    assert validate_workbook_module_manifest(payload) == []


def test_profile_contracts_filter_bindings_without_mutating_exact_cells() -> None:
    payload = _payload()
    bindings = load_json_strict(BINDING_MAP)
    full = build_profile_binding_payload(bindings, payload, resolve_module_profile(payload, "full_union"))
    core = build_profile_binding_payload(bindings, payload, resolve_module_profile(payload, "core_only"))

    original_by_id = {row["binding_id"]: row for row in bindings["bindings"]}
    assert len(full["bindings"]) == len(bindings["bindings"])
    assert len(core["bindings"]) < len(full["bindings"])
    for row in full["bindings"]:
        original = original_by_id[row["binding_id"]]
        assert row["planner_target"] == original["planner_target"]
        assert row["normalized_field"] == original["normalized_field"]
        assert row["module_id"] in full["enabled_modules"]
    assert all(row["module_id"] in core["enabled_modules"] for row in core["bindings"])

    product_pass_2a_debt_bindings = {
        row["binding_id"]
        for row in bindings["bindings"]
        if row["binding_id"].startswith("summary_revolver_availability")
        or row["binding_id"].startswith("valuation_debt_snapshot")
    }
    assert len(product_pass_2a_debt_bindings) == 30
    for profile_id in ("full_union", "anf", "pbi", "gpre"):
        resolved = resolve_module_profile(payload, profile_id)
        projected = build_profile_binding_payload(bindings, payload, resolved)
        projected_ids = {row["binding_id"] for row in projected["bindings"]}
        assert product_pass_2a_debt_bindings <= projected_ids
    core_ids = {row["binding_id"] for row in core["bindings"]}
    assert product_pass_2a_debt_bindings.isdisjoint(core_ids)


def test_module_manifest_and_profile_digests_are_contract_inputs() -> None:
    payload = _payload()
    manifest = load_json_strict(SHELL_MANIFEST)
    resolved = resolve_module_profile(payload, "full_union")
    derived = build_profile_shell_manifest(manifest, payload, resolved)

    assert derived["module_manifest"]["signature"] == canonical_json_sha256(payload)
    assert derived["module_profile"]["signature"] == canonical_json_sha256(resolved.to_dict())
    mutated = deepcopy(payload)
    mutated["modules"][0]["empty_state_behavior"] += " Changed semantics."
    assert canonical_json_sha256(mutated) != derived["module_manifest"]["signature"]


def test_hidden_support_projection_refreshes_capacity_without_renaming_shell_zone() -> None:
    payload = _payload()
    manifest = load_json_strict(SHELL_MANIFEST)
    history = next(row for row in manifest["sheets"] if row["sheet"] == "History_Q")
    history["writable_zones"][0]["zone_id"] = "established_history_contract"

    derived = build_profile_shell_manifest(manifest, payload, resolve_module_profile(payload, "full_union"))
    projected = next(row for row in derived["sheets"] if row["sheet"] == "History_Q")

    assert projected["writable_zones"][0]["zone_id"] == "established_history_contract"
    assert projected["writable_zones"][0]["target"] == "A2:G1000"
    assert projected["formulas_static_labels"] == [
        "period", "period_ordinal", "metric", "value", "unit", "source_ref", "status"
    ]


def test_formula_output_support_surfaces_are_never_binding_writable() -> None:
    payload = _payload()
    manifest = load_json_strict(SHELL_MANIFEST)
    derived = build_profile_shell_manifest(manifest, payload, resolve_module_profile(payload, "full_union"))
    sheets = {row["sheet"]: row for row in derived["sheets"]}

    assert "Valuation_Summary" not in sheets
    assert "Valuation_Grid" not in sheets
    support = sheets["{ticker}_Investment_Case_Data"]
    assert all(
        not zone["target"].startswith("BB")
        for zone in support["writable_zones"]
    )
    assert any(zone["target"] == "BB2:BQ25" for zone in support["non_writable_zones"])

    invalid = deepcopy(payload)
    investment_case_module = next(
        row for row in invalid["modules"] if row["module_id"] == "investment_case_market_implied"
    )
    formula_sheet = next(
        row for row in investment_case_module["sheets"]
        if row["sheet"] == "{ticker}_Investment_Case_Data"
    )
    formula_sheet.pop("formula_owner")
    assert any(
        "{ticker}_Investment_Case_Data" in issue and "requires a formula_owner" in issue
        for issue in validate_workbook_module_manifest(invalid)
    )


def test_user_input_zone_projection_is_idempotent() -> None:
    manifest = load_json_strict(SHELL_MANIFEST)

    _configure_investment_case_ownership_zones(manifest)
    _configure_investment_case_ownership_zones(manifest)

    investment_case = next(row for row in manifest["sheets"] if row["sheet"] == "{ticker}_Investment_Case")
    scenario_zones = [
        row for row in investment_case["writable_zones"]
        if str(row["zone_id"]).startswith("ic_scenario_user_input_")
    ]
    label_zones = [
        row for row in investment_case["non_writable_zones"]
        if str(row["zone_id"]).startswith("ic_static_label_column_")
    ]
    assert len(scenario_zones) == 27
    assert len({row["zone_id"] for row in scenario_zones}) == 27
    assert label_zones == []
    assert investment_case["rich_shell_lab_merge_floor_ratio"] == 0.10
    module_manifest = _payload()
    investment_case_block = next(
        row
        for row in module_manifest["modules"]
        if row["module_id"] == "investment_case_market_implied"
    )
    assert next(
        row["target"]
        for row in investment_case_block["visible_blocks"]
        if row["block_id"] == "investment_case"
    ) == "A1:M225"


def test_controlled_materializer_creates_isolated_profile_variant(tmp_path: Path) -> None:
    shell = tmp_path / "core-shell.xlsx"
    manifest_path = tmp_path / "core-manifest.json"
    binding_path = tmp_path / "core-bindings.json"

    materialize_shell(
        data_root=tmp_path,
        output_path=shell,
        manifest_path=SHELL_MANIFEST,
        binding_map_path=BINDING_MAP,
        module_manifest_path=MODULE_MANIFEST,
        module_profile_id="core_only",
        contract_manifest_output_path=manifest_path,
        contract_binding_map_output_path=binding_path,
        update_identity=True,
    )

    manifest = load_json_strict(manifest_path)
    bindings = load_json_strict(binding_path)
    report = verify_shell_identity(shell, manifest=manifest, binding_payload=bindings)
    wb = load_workbook(shell, read_only=False, data_only=False)
    try:
        assert report.status == "PASS", report.issues
        scenario_contracts = {
            binding_id: [
                row
                for row in manifest["planner_cell_contracts"]
                if binding_id in row.get("allowed_binding_ids", [])
            ]
            for binding_id in ("ic_bull_base_bear_rows", "ic_scenario_bridge_rows")
        }
        assert len(scenario_contracts["ic_bull_base_bear_rows"]) == 17
        assert len(scenario_contracts["ic_scenario_bridge_rows"]) == 18
        history_contract = next(row for row in manifest["sheets"] if row["sheet"] == "History_Q")
        assert history_contract["writable_zones"][0]["zone_id"] == "calculation_history_quarterly_rows"
        assert wb.sheetnames == manifest["union_sheet_order"]
        assert [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"] == manifest["visible_sheet_order"]
        assert len(wb.sheetnames) == 44
        for name, contract in sheet_contracts(_payload()).items():
            if contract["role"] != "visible_product":
                assert wb[name].sheet_state != "visible"
                header_row = int(contract.get("header_row") or 1)
                assert [
                    wb[name].cell(header_row, col).value
                    for col in range(1, len(contract["headers"]) + 1)
                ] == contract["headers"]
        resolved = resolve_module_profile(_payload(), "core_only")
        assert validate_workbook_execution_ownership(wb, _payload(), bindings, resolved) == []
        disabled_shared_cells = (
            ("SUMMARY", "A3"),
            ("SUMMARY", "A8"),
            ("SUMMARY", "A41"),
            ("BS_Segments", "A90"),
            ("Valuation", "A23"),
            ("Valuation", "A59"),
            ("Valuation", "A116"),
            ("Valuation", "A122"),
            ("Valuation", "A137"),
            ("Valuation", "O7"),
            ("Valuation", "O39"),
            ("Valuation", "O51"),
        )
        for sheet_name, coordinate in disabled_shared_cells:
            cell = wb[sheet_name][coordinate]
            assert cell.value is None, f"{sheet_name}!{coordinate} retained {cell.value!r}"
            assert cell.style_id == 0, f"{sheet_name}!{coordinate} retained style {cell.style_id}"
        assert wb["Valuation"]["A192"].value == "Forward Valuation Summary"
        assert "Valuation_Summary" not in wb.sheetnames
        assert "Valuation_Grid" not in wb.sheetnames
        for name in ("ScenarioProfile", "ScenarioImpliedPrice", "DCF_Horizon"):
            assert name not in wb.defined_names
        for name in (
            "HV_Base_MetricKey",
            "HV_Base_Value",
            "HV_Base_Status",
            "HV_Recompute_CandidateKey",
            "HV_Recompute_RowParity",
            "HV_Recompute_CandidateParity",
            "HV_Flags_CandidateKey",
            "HV_Flags_Score",
            "HV_Flags_State",
        ):
            assert name not in wb.defined_names
        assert not any(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for sheet_name in ("Hidden_Value_Audit", "Hidden_Value_Recompute", "Valuation")
            for row in wb[sheet_name].iter_rows(min_row=2)
            for cell in row
            if sheet_name != "Valuation" or 137 <= cell.row <= 143
        )
        assert all(ws.protection.sheet for ws in wb.worksheets)
        assert not any(
            cell.protection.locked is False
            for ws in wb.worksheets
            for cell in ws._cells.values()
        )
        hidden_value_bindings = {
            "hidden_value_base_rows",
            "hidden_value_audit_rows",
            "hidden_value_recompute_rows",
            "hidden_value_flags_rows",
            "hidden_value_valuation_rows",
        }
        assert all(
            row.get("planning_state") != "active"
            for row in bindings["bindings"]
            if row["binding_id"] in hidden_value_bindings
        )
    finally:
        wb.close()


def test_hidden_value_visible_ownership_is_bounded_and_support_styles_are_owner_controlled() -> None:
    payload = _payload()
    module = next(row for row in payload["modules"] if row["module_id"] == "hidden_value_signals")

    assert module["visible_blocks"] == [
        {
            "block_id": "valuation_hidden_value",
            "sheet": "Valuation",
            "target": "A137:R143",
            "empty_state": "No triggered rows; typed state counts retain the audit disposition.",
        }
    ]
    assert {(row["sheet"], row["target"]) for row in module["style_ownership"]} == {
        ("Valuation", "A137:R143"),
        ("Hidden_Value_Audit", "F2:F8"),
        ("Hidden_Value_Flags", "G2:G8"),
    }
    assert "hidden_value_valuation_rows" in module["binding_ids"]
    assert "hidden_value_valuation_state_count_formulas" in module["formula_ids"]
    assert validate_workbook_module_manifest(payload) == []

    invalid = deepcopy(payload)
    invalid_module = next(row for row in invalid["modules"] if row["module_id"] == "hidden_value_signals")
    invalid_module["style_ownership"][1]["sheet"] = "SUMMARY"
    issues = validate_workbook_module_manifest(invalid)
    assert any("neither an owned visible/profile block nor" in issue for issue in issues)


def test_profile_resolution_is_invalidated_by_profile_mutation() -> None:
    payload = _payload()
    bindings = load_json_strict(BINDING_MAP)
    resolved = resolve_module_profile(payload, "core_only")
    mutated = deepcopy(payload)
    core_profile = next(row for row in mutated["profiles"] if row["profile_id"] == "core_only")
    core_profile["enabled_modules"].append("guidance_promises")

    with pytest.raises(ValueError, match="no longer matches the manifest"):
        build_profile_binding_payload(bindings, mutated, resolved)


def test_profile_binding_payload_does_not_infer_ticker_from_prose() -> None:
    payload = _payload()
    serialized = json.dumps(payload["profiles"], ensure_ascii=False).lower()

    assert "infer" not in serialized
    assert "prose" not in serialized
    try:
        profile_id_for_ticker(payload, "UNDECLARED")
    except ValueError as exc:
        assert "no declarative workbook module profile" in str(exc)
    else:  # pragma: no cover
        raise AssertionError("An undeclared ticker acquired an inferred module profile.")
