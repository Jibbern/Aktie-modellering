from __future__ import annotations

import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from pbi_xbrl.workbook_number_formats import (
    neutralize_negative_number_format,
    neutralize_workbook_negative_number_formats,
)


def test_negative_number_format_removes_only_the_red_directive() -> None:
    assert neutralize_negative_number_format("#,##0.0;[Red]-#,##0.0") == "#,##0.0;-#,##0.0"
    assert neutralize_negative_number_format("0.0%;[RED]-0.0%") == "0.0%;-0.0%"
    assert neutralize_negative_number_format("$0.00;-$0.00") == "$0.00;-$0.00"


def test_workbook_normalization_removes_red_formats_from_cells_and_ooxml(tmp_path: Path) -> None:
    output = tmp_path / "neutral-formats.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = -1.25
    worksheet["A1"].number_format = "#,##0.0;[Red]-#,##0.0"
    worksheet["A2"] = -0.125
    worksheet["A2"].number_format = "0.0%;[Red]-0.0%"

    result = neutralize_workbook_negative_number_formats(workbook)
    workbook.save(output)
    workbook.close()

    assert result.custom_formats_changed == 2
    reopened = load_workbook(output, data_only=False, read_only=False)
    try:
        assert reopened["Sheet"]["A1"].number_format == "#,##0.0;-#,##0.0"
        assert reopened["Sheet"]["A2"].number_format == "0.0%;-0.0%"
    finally:
        reopened.close()

    with zipfile.ZipFile(output) as archive:
        styles_xml = archive.read("xl/styles.xml").decode("utf-8")
    assert "[Red]" not in styles_xml
    assert "[RED]" not in styles_xml
