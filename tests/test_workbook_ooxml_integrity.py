from __future__ import annotations

import os
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path
from zipfile import ZipFile

import openpyxl
import pytest
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


WORKBOOK_DIR = Path(
    os.environ.get(
        "STOCK_MODEL_WORKBOOK_DIR",
        r"C:\Users\Jibbe\Aktier\StockModelData\outputs\Excel stock models",
    )
)
CANONICAL_TICKERS = ("PBI", "GPRE", "ANF")

OOXML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def _workbook_path(ticker: str) -> Path:
    path = WORKBOOK_DIR / f"{ticker}_model.xlsx"
    if not path.exists():
        pytest.skip(f"Workbook not found: {path}")
    return path


def _worksheet_xml_map(path: Path) -> dict[str, str]:
    with ZipFile(path) as zf:
        workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rel_root.findall("pkg:Relationship", OOXML_NS)
        }
        out: dict[str, str] = {}
        sheets = workbook_root.find("main:sheets", OOXML_NS)
        assert sheets is not None
        for sheet in sheets.findall("main:sheet", OOXML_NS):
            name = sheet.attrib["name"]
            rel_id = sheet.attrib[
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            ]
            target = rel_targets[rel_id].lstrip("/")
            if target.startswith("xl/"):
                xml_path = target
            else:
                xml_path = f"xl/{target}"
            out[xml_path] = name
        return out


def _merged_range_issues(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[str]:
    refs = [str(merged) for merged in ws.merged_cells.ranges]
    issues: list[str] = []
    duplicates = [ref for ref, count in Counter(refs).items() if count > 1]
    for ref in duplicates:
        issues.append(f"{ws.title}: duplicate merged range {ref}")

    bounds: list[tuple[str, int, int, int, int]] = []
    for ref in refs:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref)
        except Exception as exc:  # pragma: no cover - defensive for corrupt OOXML refs
            issues.append(f"{ws.title}: invalid merged range {ref}: {exc}")
            continue
        if min_col > max_col or min_row > max_row:
            issues.append(f"{ws.title}: invalid merged range {ref}: inverted bounds")
            continue
        bounds.append((ref, min_col, min_row, max_col, max_row))

    for idx, (left_ref, left_min_col, left_min_row, left_max_col, left_max_row) in enumerate(bounds):
        for right_ref, right_min_col, right_min_row, right_max_col, right_max_row in bounds[idx + 1 :]:
            rows_overlap = left_min_row <= right_max_row and right_min_row <= left_max_row
            cols_overlap = left_min_col <= right_max_col and right_min_col <= left_max_col
            if rows_overlap and cols_overlap:
                issues.append(f"{ws.title}: overlapping merged ranges {left_ref} vs {right_ref}")
    return issues


@pytest.mark.parametrize("ticker", CANONICAL_TICKERS)
def test_workbook_sheet3_xml_maps_to_bs_segments(ticker: str) -> None:
    mapping = _worksheet_xml_map(_workbook_path(ticker))

    assert mapping.get("xl/worksheets/sheet3.xml") == "BS_Segments"


@pytest.mark.parametrize("ticker", CANONICAL_TICKERS)
def test_workbook_has_no_duplicate_invalid_or_overlapping_merged_ranges(ticker: str) -> None:
    path = _workbook_path(ticker)
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        issues = [
            issue
            for ws in wb.worksheets
            for issue in _merged_range_issues(ws)
        ]
    finally:
        wb.close()
    assert issues == []


@pytest.mark.parametrize("ticker", CANONICAL_TICKERS)
def test_bs_segments_shared_top_layout_and_standard_widths(ticker: str) -> None:
    path = _workbook_path(ticker)
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        ws = wb["BS_Segments"]
        assert str(ws["A3"].value or "").strip().startswith("QA:")
        assert str(ws["A4"].value or "").strip() == "Balance sheet & Segments"
        assert str(ws["B6"].value or "").strip() == "Actuals"
        assert str(ws["A7"].value or "").strip() == "Quarter"
        assert ws.freeze_panes in {"B8", "A8"}

        for col_idx in range(2, 10):
            letter = get_column_letter(col_idx)
            assert float(ws.column_dimensions[letter].width or 0.0) == pytest.approx(11.29, abs=0.03)
        assert float(ws.column_dimensions["A"].width or 0.0) > 11.29
    finally:
        wb.close()
