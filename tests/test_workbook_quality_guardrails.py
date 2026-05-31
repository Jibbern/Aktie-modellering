from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.styles import PatternFill

from pbi_xbrl.workbook_quality_guardrails import run_workbook_quality_guardrails


WORKBOOK_DIR = Path(
    os.environ.get(
        "STOCK_MODEL_WORKBOOK_DIR",
        r"C:\Users\Jibbe\Aktier\StockModelData\outputs\Excel stock models",
    )
)


NARRATIVE_HEADERS = [
    "Ticker",
    "Quarter",
    "Category",
    "Theme",
    "What happened",
    "Management framing",
    "Why it matters",
    "Model implication",
    "Valuation implication",
    "Double-count guardrail",
    "Linked sheet",
    "Linked metric",
    "Amount",
    "Unit",
    "Source date",
    "Source type",
    "Source / note",
    "Confidence",
    "Include in UI",
]


def _rule_ids(wb: OpenpyxlWorkbook, ticker: str) -> set[str]:
    return {issue.rule_id for issue in run_workbook_quality_guardrails(wb, ticker)}


def _promise_wb(rows: list[list[Any]]) -> OpenpyxlWorkbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Promise_Progress_UI"
    ws.append(["2025-Q4 revisions"])
    ws.append(
        [
            "Metric",
            "Previous guide",
            "New/current guide",
            "Change type",
            "Actual",
            "Progress / run-rate",
            "Status",
            "Horizon",
            "Stated in",
            "Source date",
            "Source / note",
            "",
            "",
            "",
            "",
        ]
    )
    for row in rows:
        ws.append(row)
    return wb


def test_promise_guardrails_detect_horizon_q4_and_duplicate_regressions() -> None:
    wb = _promise_wb(
        [
            [
                "Revenue guidance",
                "",
                "$100m",
                "Initial",
                "$25m",
                "",
                "Completed",
                "2026 year",
                "2025-Q4",
                "2026-02-01",
                "2026 year guidance.",
                "",
                "",
                "",
                "guidance:revenue_guidance:2026_year",
            ],
            [
                "Revenue guidance",
                "",
                "$110m",
                "Updated",
                "",
                "",
                "Open",
                "2026 year",
                "2025-Q4",
                "2026-02-01",
                "Duplicate future-year row.",
                "",
                "",
                "",
                "guidance:revenue_guidance:2026_year",
            ],
            [
                "Adjusted EBIT guidance",
                "",
                "$450m",
                "Updated",
                "FY: $461.3m",
                "",
                "Hit",
                "2025 year",
                "2025-Q4",
                "2026-02-01",
                "Final annual row with FY value in Actual.",
                "",
                "",
                "",
                "guidance:adjusted_ebit_guidance:2025_year",
            ],
            [
                "FCF target",
                "",
                "$300m",
                "Updated",
                "$70m",
                "YTD: $200m",
                "On track",
                "2025 year",
                "2025-Q3",
                "2025-11-01",
                "Moved backward from the section event.",
                "",
                "",
                "",
                "guidance:fcf_target:2025_year",
            ],
        ]
    )

    assert {
        "promise_future_annual_in_prior_year_q4",
        "promise_duplicate_metric_horizon",
        "promise_q4_actual_progress_split",
        "promise_stated_in_mismatch",
    }.issubset(_rule_ids(wb, "PBI"))


def test_promise_hidden_key_guardrails_require_source_backed_alignment() -> None:
    wb = _promise_wb(
        [
            [
                "Adjusted EPS",
                "",
                "$10.20-$11.00",
                "Initial",
                "",
                "",
                "Open",
                "2026 year",
                "2025-Q4",
                "2026-03-04",
                "Earnings release source-backed guidance.",
                "",
                "",
                "",
                "",
            ],
            [
                "Operating margin",
                "",
                "12.0-12.5%",
                "Initial",
                "",
                "",
                "Open",
                "2026 year",
                "2025-Q4",
                "2026-03-04",
                "Earnings release source-backed guidance.",
                "",
                "",
                "",
                "guidance:capex:2026_year:2026_q1:2026_03_04",
            ],
            [
                "Scenario placeholder",
                "",
                "Manual assumption",
                "Initial",
                "",
                "",
                "Open",
                "2025-Q4",
                "2025-Q4",
                "",
                "manual no-source scenario assumption",
                "",
                "",
                "",
                "",
            ],
        ]
    )

    issues = run_workbook_quality_guardrails(wb, "ANF")
    rule_ids = {issue.rule_id for issue in issues}

    assert "promise_source_backed_missing_hidden_key" in rule_ids
    assert "promise_hidden_key_metric_mismatch" in rule_ids
    assert all(issue.metric_label != "Scenario placeholder" for issue in issues)


def test_quarter_narrative_amount_guardrails_detect_descriptor_prose() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quarter_Narrative_Data"
    ws.append(NARRATIVE_HEADERS)
    ws.append(["ANF", "2026-Q1", "", "Brand cuts", "", "", "", "", "", "", "", "", "Americas are geographic segments, not value text.", "", "", "", "", "", "Yes"])
    ws.append(["ANF", "2026-Q1", "", "Context", "", "", "", "", "", "", "", "", "This amount field contains long context words that should live in narrative columns instead", "", "", "", "", "", "Yes"])
    ws.append(["ANF", "2026-Q1", "", "Tariff", "", "", "", "", "", "", "", "", "-290 / -70", "bps", "", "", "", "", "Yes"])

    issues = run_workbook_quality_guardrails(wb, "ANF")
    by_rule = {issue.rule_id: issue for issue in issues}

    assert by_rule["narrative_amount_descriptor_prose"].severity == "P1"
    assert by_rule["narrative_amount_long_prose"].severity == "P2"


def test_sector_specific_guardrails_catch_blank_or_missing_carbon_rows() -> None:
    pbi = Workbook()
    pbi.active.title = "BS_Segments"
    pbi["BS_Segments"].append(["Metric", "2026-Q1"])
    pbi["BS_Segments"].append(["Carbon equipment liabilities", ""])

    gpre_missing = Workbook()
    gpre_missing.active.title = "BS_Segments"
    gpre_missing["BS_Segments"].append(["Metric", "2026-Q1"])
    gpre_missing["BS_Segments"].append(["Total assets", 1.0])

    gpre_ok = Workbook()
    gpre_ok.active.title = "BS_Segments"
    gpre_ok["BS_Segments"].append(["Metric", "2026-Q1"])
    gpre_ok["BS_Segments"].append(["Carbon equipment liabilities", 12.3])

    assert "sector_blank_carbon_row" in _rule_ids(pbi, "PBI")
    assert "sector_carbon_row_missing" in _rule_ids(gpre_missing, "GPRE")
    assert not run_workbook_quality_guardrails(gpre_ok, "GPRE")


def test_hidden_value_guardrails_match_flags_to_valuation_display() -> None:
    wb = Workbook()
    flags = wb.active
    flags.title = "Hidden_Value_Flags"
    flags.append(
        [
            "rank",
            "flag_code",
            "title",
            "score",
            "severity",
            "as_of_quarter",
            "evidence_1",
            "evidence_2",
            "evidence_3",
            "metrics_json",
            "visible_support",
            "triggered",
        ]
    )
    flags.append([1, "F", "Share count reduction", 100, "High", "2026-03-31", "", "", "", "{}", "Shares down", 1])
    flags.append([2, "A", "Non-triggered EBIT row", 20, "Info", "2026-03-31", "", "", "", "{}", "Not active", 0])
    flags.append([3, "C", "Cashflow quality", 55, "Med", "2026-03-31", "", "", "", '{"fcf_yield": null}', "(price-linked)", 1])
    flags.append([4, "G", "Incomplete row", "", "Info", "2026-03-31", "", "", "", "{}", "", ""])

    valuation = wb.create_sheet("Valuation")
    valuation.append(["Hidden value flags"])
    valuation.append(["Flag", "Summary", "", "", "", "Score", "Severity", "Result / support"])
    valuation.append(["Flag 1", "Non-triggered EBIT row", "", "", "", 20, "Info", "Not active"])

    assert {
        "hidden_value_triggered_missing_from_valuation",
        "hidden_value_nontriggered_leaked_to_valuation",
        "hidden_value_price_linked_trigger_without_inputs",
        "hidden_value_score_or_trigger_blank",
    }.issubset(_rule_ids(wb, "PBI"))


def test_hidden_value_display_scan_uses_labels_not_fixed_row_offsets() -> None:
    wb = Workbook()
    flags = wb.active
    flags.title = "Hidden_Value_Flags"
    flags.append(
        [
            "rank",
            "flag_code",
            "title",
            "score",
            "severity",
            "as_of_quarter",
            "evidence_1",
            "evidence_2",
            "evidence_3",
            "metrics_json",
            "visible_support",
            "triggered",
        ]
    )
    flags.append([1, "F", "Share count reduction", 100, "High", "2026-03-31", "", "", "", "{}", "Shares down", 1])

    valuation = wb.create_sheet("Valuation")
    valuation.append(["Hidden value flags"])
    valuation.append(["Flag", "Summary", "", "", "", "Score", "Severity", "Result / support"])
    for _ in range(8):
        valuation.append(["", "", "", "", "", "", "", ""])
    valuation.append(["Flag 1", "Share count reduction", "", "", "", 100, "High", "Shares down"])
    valuation.append(["Operating signals"])

    assert "hidden_value_triggered_missing_from_valuation" not in _rule_ids(wb, "PBI")


def _debt_current_source_wb(visible_value: Any) -> OpenpyxlWorkbook:
    wb = Workbook()
    bs = wb.active
    bs.title = "BS_Segments"
    bs.append(["Balance sheet & Segments"])
    bs.append(["Quarter", "2026-Q1"])
    bs.append(["Current maturities of long-term debt", visible_value])

    facts = wb.create_sheet("DATA_Facts_Long")
    facts.append(["metric", "period_end", "period_type", "value", "unit", "source_class", "method", "qa_severity"])
    facts.append(["debt_current", datetime(2026, 3, 31), "Instant", 69_316_000, "USD", "xbrl_fact", "direct", "PASS"])
    return wb


def test_source_backed_bs_segments_flags_blank_debt_current() -> None:
    issues = run_workbook_quality_guardrails(_debt_current_source_wb(""), "GPRE")

    matches = [issue for issue in issues if issue.rule_id == "source_backed_missing_bs_segment_value"]
    assert len(matches) == 1
    assert matches[0].severity == "P1"
    assert matches[0].sheet == "BS_Segments"
    assert matches[0].metric_label == "Current maturities of long-term debt"


def test_source_backed_bs_segments_allows_populated_debt_current() -> None:
    assert "source_backed_missing_bs_segment_value" not in _rule_ids(_debt_current_source_wb(69.316), "GPRE")


def test_source_backed_promise_flags_blank_actual_or_progress_from_hidden_audit() -> None:
    hidden_key = "guidance:45z_monetization:2025_q4"
    wb = _promise_wb(
        [
            [
                "45Z monetization",
                None,
                "quarter-specific disclosure",
                "Initial",
                "",
                "",
                "Open",
                "2025-Q4",
                "2025-Q4",
                "2025-12-31",
                "source-backed earnings release row.",
                "",
                "",
                "",
                hidden_key,
            ]
        ]
    )
    audit = wb.create_sheet("Promise_Source_Audit")
    audit.append(["hidden_key", "actual", "progress", "source_date"])
    audit.append([hidden_key, "$23.4m", "YTD: $49.9m", "2025-12-31"])

    issues = run_workbook_quality_guardrails(wb, "GPRE")

    assert "source_backed_missing_promise_visible_value" in {issue.rule_id for issue in issues}


def test_source_backed_operating_drivers_flags_blank_direct_bs_segment_map() -> None:
    wb = Workbook()
    bs = wb.active
    bs.title = "BS_Segments"
    bs.append(["Balance sheet & Segments"])
    bs.append(["Quarter", "2026-Q1"])
    bs.append(["Quarterly segments"])
    bs.append(["Revenue"])
    bs.append(["Presort Services", 163.466])

    od = wb.create_sheet("Operating_Drivers")
    od.append(["Operating Drivers"])
    od.append(["Segment support - latest 12 quarters"])
    od.append(["Metric / segment", "2026-Q1"])
    od.append(["Revenue"])
    od.append(["Presort Services", ""])

    issues = run_workbook_quality_guardrails(wb, "PBI")

    assert "source_backed_missing_operating_driver_value" in {issue.rule_id for issue in issues}


def test_comparison_coloring_flags_clean_comparator_with_neutral_fill() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation"
    ws.append(["Quarter", "2025-Q1", "2026-Q1"])
    ws.append(["Revenue", 100.0, 125.0])

    issues = run_workbook_quality_guardrails(wb, "ANF")
    matches = [issue for issue in issues if issue.rule_id == "comparison_coloring_clean_comparator_neutral"]

    assert len(matches) == 1
    assert matches[0].severity == "P2"
    assert matches[0].sheet == "Valuation"
    assert matches[0].metric_label == "Revenue"


def test_comparison_coloring_flags_colored_cell_with_tiny_comparator() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Operating_Drivers"
    ws.append(["Quarter", "2025-Q1", "2026-Q1"])
    ws.append(["Operating margin %", 0.0, 0.14])
    ws.cell(2, 3).fill = PatternFill("solid", fgColor="2F80ED")

    issues = run_workbook_quality_guardrails(wb, "ANF")
    matches = [issue for issue in issues if issue.rule_id == "comparison_coloring_without_clean_comparator"]

    assert len(matches) == 1
    assert matches[0].severity == "P2"
    assert matches[0].sheet == "Operating_Drivers"
    assert matches[0].metric_label == "Operating margin %"


def test_generated_workbooks_have_no_blocking_quality_guardrail_issues() -> None:
    for ticker in ("PBI", "GPRE", "ANF"):
        path = next(
            (
                WORKBOOK_DIR / f"{ticker}_model{suffix}"
                for suffix in (".xlsm", ".xlsx")
                if (WORKBOOK_DIR / f"{ticker}_model{suffix}").exists()
            ),
            WORKBOOK_DIR / f"{ticker}_model.xlsx",
        )
        if not path.exists():
            pytest.skip(f"{path} is not available for workbook quality guardrail readback")
        wb = load_workbook(path, data_only=False, read_only=False, keep_vba=True)
        try:
            blocking = [
                issue
                for issue in run_workbook_quality_guardrails(wb, ticker)
                if issue.severity in {"P0", "P1"}
            ]
            assert not blocking, [issue.to_dict() for issue in blocking]
        finally:
            wb.close()
