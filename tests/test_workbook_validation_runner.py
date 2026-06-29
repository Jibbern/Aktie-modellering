from __future__ import annotations

import csv

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName

from pbi_xbrl.workbook_validation_runner import (
    BAD_MARKER_TERMS,
    TICKERS,
    ValidationConfig,
    default_workbook_paths,
    resolve_workbook_paths,
    summary_rows,
    validate_workbook,
    validate_workbooks,
    write_validation_reports,
)


def _add_required_sheets(wb: Workbook, ticker: str) -> None:
    if wb.active is not None:
        wb.active.title = "Valuation"
    for sheet_name in [
        f"{ticker}_Investment_Case",
        "Promise_Progress_UI",
        "Quarter_Notes_UI",
        "History_Q",
        "Operating_Drivers",
        "Needs_Review",
        "QA_Log",
        "QA_Checks",
        "Scenario_Bridge_Tax_Treatment",
        "Scenario_Driver_Assumptions",
        "Quarter_Narrative_Data",
        "BS_Segments",
    ]:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)


def _add_required_named_ranges(wb: Workbook) -> None:
    wb["Valuation"]["A1"] = 0.12
    for name in [
        "CompanyOperatingMargin_Latest",
        "OperatingMargin_Latest",
        "CompanyOperatingMargin_TTM",
    ]:
        wb.defined_names.add(DefinedName(name=name, attr_text="'Valuation'!$A$1"))


def _make_clean_validation_workbook(path: Path, ticker: str = "PBI") -> Path:
    wb = Workbook()
    _add_required_sheets(wb, ticker)
    _add_required_named_ranges(wb)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb["Needs_Review"].append(["priority", "issue"])
    wb["Needs_Review"].append(["P2", "allowed non-P1 note"])
    wb["QA_Log"].append(["check", "status"])
    wb["QA_Log"].append(["visible status", "pass"])
    wb["Promise_Progress_UI"]["A1"] = "2025-Q4"
    wb["Quarter_Notes_UI"]["A1"] = "2025-Q4 - Quarter Notes"
    if ticker == "GPRE":
        wb["BS_Segments"].append(["Metric", "2026-Q1"])
        wb["BS_Segments"].append(["Carbon equipment liabilities", 12.3])
    wb.save(path)
    return path


def test_validation_runner_passes_clean_workbook_and_writes_reports(tmp_path: Path) -> None:
    workbook_path = _make_clean_validation_workbook(tmp_path / "PBI_model.xlsx", "PBI")

    result = validate_workbook(workbook_path, "PBI")

    assert result.overall == "PASS"
    assert result.formula_error_count == 0
    assert result.needs_review_p1_count == 0
    assert result.qa_blank_nan_status_count == 0
    assert result.cross_company_leakage_count == 0
    assert result.bad_marker_count == 0
    assert result.missing_required_sheets == []
    assert result.missing_named_ranges == []
    assert result.calc_settings_ok
    assert result.elapsed_seconds >= 0.0

    report_paths = write_validation_reports([result], tmp_path / "validation")
    assert report_paths["json"].exists()
    assert report_paths["csv"].exists()
    assert report_paths["guardrails_json"].exists()
    assert report_paths["guardrails_csv"].exists()
    row = summary_rows([result])[0]
    assert "Skipped large sheets" in row
    assert "Sampled sheets" in row
    assert "Guardrail P0/P1" in row
    assert "Guardrail P2" in row
    assert "Elapsed seconds" in row


def test_validation_runner_reports_cells_and_values_for_failures(tmp_path: Path) -> None:
    workbook_path = _make_clean_validation_workbook(tmp_path / "PBI_model.xlsx", "PBI")
    wb = Workbook()
    _add_required_sheets(wb, "PBI")
    wb.calculation.calcMode = "manual"
    wb["Valuation"]["A1"] = "#REF!"
    wb["PBI_Investment_Case"]["A1"] = "45Z should not leak into PBI"
    wb["Promise_Progress_UI"]["A1"] = "Actual / latest actual"
    wb["Promise_Progress_UI"]["B1"] = "Q4 FY2025"
    wb["Quarter_Notes_UI"]["A1"] = "DEBUG note"
    wb["Needs_Review"].append(["priority", "issue"])
    wb["Needs_Review"].append(["P1", "critical issue"])
    wb["QA_Log"].append(["check", "status"])
    wb["QA_Log"].append(["bad status", "nan"])
    wb.save(workbook_path)

    result = validate_workbook(workbook_path, "PBI")
    details = "\n".join(issue.detail for issue in result.issues)

    assert result.overall == "FAIL"
    assert result.formula_error_count == 1
    assert result.needs_review_p1_count == 1
    assert result.qa_blank_nan_status_count == 1
    assert result.cross_company_leakage_count == 1
    assert result.bad_marker_count >= 2
    assert result.quarter_label_issue_count == 1
    assert set(result.missing_named_ranges) == {
        "CompanyOperatingMargin_Latest",
        "OperatingMargin_Latest",
        "CompanyOperatingMargin_TTM",
    }
    assert not result.calc_settings_ok
    assert "Valuation!A1" in details
    assert "PBI_Investment_Case!A1" in details
    assert "Promise_Progress_UI!A1" in details
    assert "QA_Log!B2" in details


def test_validation_runner_batches_all_tickers(tmp_path: Path) -> None:
    paths = {
        "PBI": _make_clean_validation_workbook(tmp_path / "PBI_model.xlsx", "PBI"),
        "GPRE": _make_clean_validation_workbook(tmp_path / "GPRE_model.xlsx", "GPRE"),
        "ANF": _make_clean_validation_workbook(tmp_path / "ANF_model.xlsx", "ANF"),
    }

    results = validate_workbooks(paths)

    assert [result.ticker for result in results] == ["PBI", "GPRE", "ANF"]
    assert all(result.overall == "PASS" for result in results)


def test_validation_runner_uses_explicit_xlsx_workbook_path(tmp_path: Path) -> None:
    explicit_path = _make_clean_validation_workbook(tmp_path / "custom_pbi_snapshot.xlsx", "PBI")
    stale_default = _make_clean_validation_workbook(tmp_path / "PBI_model.xlsx", "PBI")
    wb = Workbook()
    wb.active.title = "Valuation"
    wb["Valuation"]["A1"] = "#REF!"
    wb.save(stale_default)

    paths = resolve_workbook_paths(workbook_dir=explicit_path, tickers=["PBI"])
    assert paths == {"PBI": explicit_path.resolve()}

    result = validate_workbooks(paths)[0]
    assert result.path == str(explicit_path.resolve())
    assert result.overall == "PASS"


def test_validation_runner_uses_explicit_xlsm_workbook_path(tmp_path: Path) -> None:
    explicit_path = _make_clean_validation_workbook(tmp_path / "GPRE_model.xlsm", "GPRE")
    stale_xlsx = _make_clean_validation_workbook(tmp_path / "GPRE_model.xlsx", "GPRE")
    wb = Workbook()
    wb.active.title = "Valuation"
    wb["Valuation"]["A1"] = "#VALUE!"
    wb.save(stale_xlsx)

    paths = resolve_workbook_paths(workbook_dir=explicit_path, tickers=["GPRE"])
    assert paths == {"GPRE": explicit_path.resolve()}

    result = validate_workbooks(paths)[0]
    assert result.path == str(explicit_path.resolve())
    assert result.overall == "PASS"


def test_validation_runner_defaults_include_gtx_and_accept_xlsx_only(tmp_path: Path) -> None:
    assert "GTX" in TICKERS

    explicit_path = _make_clean_validation_workbook(tmp_path / "GTX_model.xlsx", "GTX")
    defaults = default_workbook_paths(tmp_path)

    assert defaults["GTX"] == explicit_path

    paths = resolve_workbook_paths(workbook_dir=explicit_path, tickers=["GTX"])

    assert paths == {"GTX": explicit_path.resolve()}
    results = validate_workbooks(paths)
    assert [result.ticker for result in results] == ["GTX"]
    assert results[0].path == str(explicit_path.resolve())
    assert results[0].overall == "PASS"


def test_validation_runner_folder_mode_prefers_existing_xlsm_outputs(tmp_path: Path) -> None:
    _make_clean_validation_workbook(tmp_path / "PBI_model.xlsm", "PBI")
    paths = default_workbook_paths(tmp_path)

    assert paths["PBI"].name == "PBI_model.xlsm"
    assert paths["GPRE"].name == "GPRE_model.xlsx"


def test_validation_runner_missing_explicit_workbook_path_reports_exact_path(tmp_path: Path) -> None:
    missing_path = tmp_path / "ANF_custom_missing.xlsm"

    paths = resolve_workbook_paths(workbook_dir=missing_path, tickers=["ANF"])
    assert paths == {"ANF": missing_path.resolve()}

    result = validate_workbooks(paths)[0]
    assert result.overall == "FAIL"
    assert result.issues[0].category == "workbook_missing"
    assert str(missing_path.resolve()) in result.issues[0].detail


def test_validation_runner_samples_large_raw_sheets(tmp_path: Path) -> None:
    workbook_path = _make_clean_validation_workbook(tmp_path / "PBI_model.xlsx", "PBI")
    wb = Workbook()
    _add_required_sheets(wb, "PBI")
    _add_required_named_ranges(wb)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb["Needs_Review"].append(["priority", "issue"])
    wb["Needs_Review"].append(["P2", "allowed non-P1 note"])
    wb["QA_Log"].append(["check", "status"])
    wb["QA_Log"].append(["visible status", "pass"])
    raw = wb.create_sheet("economics_market_raw")
    raw.append(["date", "series", "value"])
    raw.append(["2025-01-01", "head", 1])
    raw.append(["2025-01-02", "#REF! hidden inside large raw sample gap", 2])
    raw.append(["2025-01-03", "tail", 3])
    wb.save(workbook_path)

    result = validate_workbook(
        workbook_path,
        "PBI",
        config=ValidationConfig(huge_sheet_row_threshold=2, sample_head_rows=1, sample_tail_rows=1),
    )

    assert result.overall == "PASS"
    assert result.formula_error_count == 0
    assert "economics_market_raw" in result.skipped_large_sheets
    assert "economics_market_raw" in result.sampled_sheets


def test_bad_marker_terms_include_user_requested_regression_strings() -> None:
    assert "Base active values" in BAD_MARKER_TERMS
    assert "Actual / latest actual" in BAD_MARKER_TERMS
    assert "Separate revenue cut; not summed" in BAD_MARKER_TERMS


def test_validation_runner_fails_on_blocking_quality_guardrails(tmp_path: Path) -> None:
    workbook_path = _make_clean_validation_workbook(tmp_path / "PBI_model.xlsx", "PBI")
    wb = load_workbook(workbook_path)
    ws = wb["Promise_Progress_UI"]
    ws.append(["2025-Q4 revisions"])
    ws.append(
        [
            "Metric",
            "Previous guide",
            "New/current guide",
            "Change type",
            "Actual",
            "Progress / run-rate",
            "Status",
            "Horizon",
            "Stated in",
            "Source date",
            "Source / note",
            "",
            "",
            "",
            "",
        ]
    )
    ws.append(
        [
            "Revenue guidance",
            "",
            "$100m",
            "Initial",
            "$25m",
            "",
            "Completed",
            "2026 year",
            "2025-Q4",
            "2026-02-01",
            "2026 year guidance.",
            "",
            "",
            "",
            "guidance:revenue_guidance:2026_year",
        ]
    )
    wb.save(workbook_path)

    result = validate_workbook(workbook_path, "PBI")

    assert result.overall == "FAIL"
    assert result.quality_guardrail_p0_p1_count >= 1
    assert any(
        issue["rule_id"] == "promise_future_annual_in_prior_year_q4"
        for issue in result.quality_guardrail_issues
    )
    guardrail_issue = next(
        issue
        for issue in result.quality_guardrail_issues
        if issue["rule_id"] == "promise_future_annual_in_prior_year_q4"
    )
    assert guardrail_issue["guardrail_name"] == "promise_future_annual_in_prior_year_q4"
    assert guardrail_issue["workbook_path"] == str(workbook_path)
    assert guardrail_issue["failure_area"] == "model correctness"
    assert guardrail_issue["owner"] == "Promise_Progress_UI horizon routing"
    assert any(issue.category == "quality_guardrail_p1" for issue in result.issues)

    report_paths = write_validation_reports([result], tmp_path / "validation")
    with report_paths["guardrails_csv"].open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert rows
    assert rows[0]["guardrail_name"] == "promise_future_annual_in_prior_year_q4"
    assert rows[0]["workbook_path"] == str(workbook_path)
    assert rows[0]["failure_area"] == "model correctness"


def test_validation_runner_can_warn_only_on_blocking_quality_guardrails(tmp_path: Path) -> None:
    workbook_path = _make_clean_validation_workbook(tmp_path / "PBI_model.xlsx", "PBI")
    wb = load_workbook(workbook_path)
    ws = wb["Promise_Progress_UI"]
    ws.append(["2025-Q4 revisions"])
    ws.append(
        [
            "Metric",
            "Previous guide",
            "New/current guide",
            "Change type",
            "Actual",
            "Progress / run-rate",
            "Status",
            "Horizon",
            "Stated in",
            "Source date",
            "Source / note",
            "",
            "",
            "",
            "",
        ]
    )
    ws.append(
        [
            "Revenue guidance",
            "",
            "$100m",
            "Initial",
            "$25m",
            "",
            "Completed",
            "2026 year",
            "2025-Q4",
            "2026-02-01",
            "2026 year guidance.",
            "",
            "",
            "",
            "guidance:revenue_guidance:2026_year",
        ]
    )
    wb.save(workbook_path)

    result = validate_workbook(
        workbook_path,
        "PBI",
        config=ValidationConfig(quality_guardrails_warn_only=True),
    )

    assert result.overall == "PASS"
    assert result.quality_guardrail_p0_p1_count >= 1
    assert result.quality_guardrails_warn_only is True


def test_validation_runner_reports_p2_quality_guardrails_without_failing(tmp_path: Path) -> None:
    workbook_path = _make_clean_validation_workbook(tmp_path / "PBI_model.xlsx", "PBI")
    wb = load_workbook(workbook_path)
    ws = wb["Quarter_Narrative_Data"]
    headers = [
        "Ticker",
        "Quarter",
        "Category",
        "Theme",
        "What happened",
        "Management framing",
        "Why it matters",
        "Model implication",
        "Valuation implication",
        "Double-count guardrail",
        "Linked sheet",
        "Linked metric",
        "Amount",
        "Unit",
        "Source date",
        "Source type",
        "Source / note",
        "Confidence",
        "Include in UI",
    ]
    for cc, header in enumerate(headers, start=1):
        ws.cell(1, cc, header)
    ws.append(
        [
            "PBI",
            "2026-Q1",
            "Context",
            "Context row",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "This amount field contains long context words that should stay out of the value column",
            "",
            "",
            "",
            "",
            "medium",
            "Yes",
        ]
    )
    wb.save(workbook_path)

    result = validate_workbook(workbook_path, "PBI")

    assert result.overall == "PASS"
    assert result.quality_guardrail_p0_p1_count == 0
    assert result.quality_guardrail_p2_count == 1
    assert result.quality_guardrail_issues[0]["rule_id"] == "narrative_amount_long_prose"
    assert result.quality_guardrail_issues[0]["failure_area"] == "intentional exception missing"
